#!/usr/bin/env python3
"""
Generate Kit Gestión de Personal y Turnos (9 archivos: 7 main + 2 bonus).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-gestion-personal"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Shift colors
MANANA_COLOR = "FFF9C4"      # Yellow - morning
TARDE_COLOR = "BBDEFB"       # Blue - afternoon
NOCHE_COLOR = "E1BEE7"       # Purple - night
PARTIDO_COLOR = "FFE0B2"     # Orange - split shift
LIBRE_COLOR = "C8E6C9"       # Green - day off
VACACIONES_COLOR = "90CAF9"  # Blue - vacation
BAJA_COLOR = "EF9A9A"        # Red - sick leave
FESTIVO_COLOR = "A5D6A7"     # Green - holiday
PERMISO_COLOR = "FFCC80"     # Orange - leave

# Semaphore colors
VERDE_COLOR = "C8E6C9"
AMARILLO_COLOR = "FFF9C4"
ROJO_COLOR = "FFCDD2"

# Category colors
DOC_COLOR = "E3F2FD"
FORMACION_COLOR = "FFF3E0"
EQUIPO_COLOR = "E8F5E9"
OPERATIVA_COLOR = "F3E5F5"
PRUEBA_COLOR = "FFEBEE"
INPUT_COLOR = "E8F5E9"

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)
formula_font = Font(name="Calibri", size=11, color="1565C0", bold=True)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

BRAND_LINE = "AI Chef Pro · aichef.pro — Kit Gestión de Personal y Turnos"


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def style_header_row(ws, row, headers, col_start=1):
    for col_idx, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def add_brand_footer(ws, row, merge_end="G"):
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


def set_col_widths(ws, widths):
    """widths is a dict {col_letter: width} or list of (letter, width)."""
    if isinstance(widths, dict):
        widths = widths.items()
    for letter, w in widths:
        ws.column_dimensions[letter].width = w


# ═══════════════════════════════════════════════════════════
# 01 — CUADRANTE DE TURNOS SEMANAL
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Cuadrante de Turnos Semanal", [
        "Cómo usar esta plantilla:",
        "▸ Introduce los nombres de tus empleados en la columna A.",
        "▸ Asigna turnos por código: M (mañana), T (tarde), N (noche), P (partido), L (libre).",
        "▸ Las horas se calculan automáticamente en la columna 'Total Horas'.",
        "▸ Código de colores: amarillo=mañana, azul=tarde, morado=noche, naranja=partido, verde=libre.",
        "",
        "Alertas automáticas:",
        "▸ La columna 'Alertas' marca en ROJO si un empleado supera 40h semanales.",
        "▸ Recuerda: mínimo 11h de descanso entre turnos (Estatuto Trabajadores Art. 34).",
        "▸ Máximo 9h de jornada ordinaria diaria (Art. 34.3 ET).",
        "",
        "Turnos estándar:",
        "▸ M = Mañana (7:00-15:00) → 8h",
        "▸ T = Tarde (15:00-23:00) → 8h",
        "▸ N = Noche (23:00-7:00) → 8h",
        "▸ P = Partido (10:00-15:00 + 19:00-23:00) → 9h",
        "▸ L = Libre → 0h",
    ])

    # --- Sheet 2: Cuadrante Semanal ---
    ws = wb.create_sheet("Cuadrante Semanal")
    ws.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws, [("A", 22), ("B", 14), ("C", 14), ("D", 14), ("E", 14),
                        ("F", 14), ("G", 14), ("H", 14), ("I", 14), ("J", 16)])

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Cuadrante de Turnos — Semana del ___/___/___ al ___/___/___"
    ws["A1"].font = title_font
    ws.merge_cells("A2:J2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    # Shift legend
    row = 3
    legend_data = [("M=Mañana", MANANA_COLOR), ("T=Tarde", TARDE_COLOR),
                   ("N=Noche", NOCHE_COLOR), ("P=Partido", PARTIDO_COLOR), ("L=Libre", LIBRE_COLOR)]
    for col_idx, (txt, color) in enumerate(legend_data, 2):
        cell = ws.cell(row=row, column=col_idx, value=txt)
        cell.font = Font(name="Calibri", size=9, bold=True)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = center_align

    headers = ["Empleado", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo", "Total Horas", "Alertas"]
    row = 5
    style_header_row(ws, row, headers)

    # Shift validation
    dv_shift = DataValidation(type="list", formula1='"M,T,N,P,L"', allow_blank=True)
    dv_shift.error = "Usa M, T, N, P o L"
    dv_shift.prompt = "M=Mañana, T=Tarde, N=Noche, P=Partido, L=Libre"
    ws.add_data_validation(dv_shift)

    # 15 employee rows
    for i in range(15):
        r = 6 + i
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=1).font = data_font
        ws.cell(row=r, column=1).fill = input_fill

        for c in range(2, 9):  # Mon-Sun columns B-H
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = data_font
            cell.fill = input_fill
            dv_shift.add(cell)

        # Total hours formula: M=8, T=8, N=8, P=9, L=0
        col_b = get_column_letter(2)
        col_h = get_column_letter(8)
        formula_parts = []
        for c in range(2, 9):
            cl = get_column_letter(c)
            formula_parts.append(f'IF({cl}{r}="M",8,IF({cl}{r}="T",8,IF({cl}{r}="N",8,IF({cl}{r}="P",9,0))))')
        total_formula = "=" + "+".join(formula_parts)
        total_cell = ws.cell(row=r, column=9, value=total_formula)
        total_cell.font = formula_font
        total_cell.alignment = center_align
        total_cell.border = thin_border

        # Alert formula: >40h = RED warning
        alert_cell = ws.cell(row=r, column=10)
        alert_cell.value = f'=IF(I{r}>40,"⚠ EXCESO",IF(I{r}>0,"OK",""))'
        alert_cell.font = data_font
        alert_cell.alignment = center_align
        alert_cell.border = thin_border

    # Totals row
    r_total = 21
    ws.cell(row=r_total, column=1, value="TOTAL HORAS EQUIPO").font = bold_font
    ws.cell(row=r_total, column=1).border = thin_border
    ws.cell(row=r_total, column=9, value=f"=SUM(I6:I20)").font = formula_font
    ws.cell(row=r_total, column=9).alignment = center_align
    ws.cell(row=r_total, column=9).border = thin_border

    add_brand_footer(ws, 23, "J")

    # --- Sheet 3: Cuadrante Mensual ---
    ws2 = wb.create_sheet("Cuadrante Mensual")
    ws2.sheet_properties.tabColor = "0D47A1"
    set_col_widths(ws2, [("A", 22)])

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "Cuadrante Mensual — Mes: _______________"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:F2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    for week_num in range(1, 5):
        start_row = 3 + (week_num - 1) * 19
        ws2.cell(row=start_row, column=1, value=f"SEMANA {week_num}").font = section_font
        headers_m = ["Empleado", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo", "Horas"]
        style_header_row(ws2, start_row + 1, headers_m)
        for col_idx in range(2, 10):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 12
        for i in range(15):
            r = start_row + 2 + i
            for c in range(1, 10):
                cell = ws2.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = center_align
                if c == 1:
                    cell.fill = input_fill

    add_brand_footer(ws2, 80, "I")

    fname = os.path.join(OUTPUT_DIR, "01-cuadrante-turnos-semanal.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# 02 — CONTROL DE HORAS EXTRAS
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Control de Horas Extras", [
        "Cómo usar esta plantilla:",
        "▸ Registra cada día las horas de entrada y salida de cada empleado.",
        "▸ Las horas trabajadas y las horas extra se calculan automáticamente.",
        "▸ El Resumen Mensual agrega por empleado el total de horas extra y su coste.",
        "",
        "Legislación española (Estatuto de los Trabajadores):",
        "▸ Máximo 80 horas extra al año por empleado.",
        "▸ Las primeras 80h se pagan ×1.75 del salario/hora (o según convenio).",
        "▸ Horas por encima de 80 anuales: ×2.0 (o compensación en descanso).",
        "▸ Las horas extra son voluntarias salvo pacto en convenio o contrato.",
        "",
        "Columnas del registro:",
        "▸ Horas Contratadas = jornada diaria del empleado (ej: 8h).",
        "▸ Horas Extra = max(0, horas trabajadas - horas contratadas).",
        "▸ Tipo: Voluntaria / Obligatoria (fuerza mayor).",
    ])

    # --- Sheet 2: Registro Horas ---
    ws = wb.create_sheet("Registro Horas")
    ws.sheet_properties.tabColor = "E65100"
    set_col_widths(ws, [("A", 22), ("B", 14), ("C", 14), ("D", 14), ("E", 16),
                        ("F", 14), ("G", 16), ("H", 18)])

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Registro de Horas — Mes: _______________"
    ws["A1"].font = title_font
    ws.merge_cells("A2:H2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Empleado", "Fecha", "Hora Entrada", "Hora Salida", "Horas Trabajadas", "H. Contratadas", "Horas Extra", "Tipo Extra"]
    style_header_row(ws, 4, headers)

    dv_tipo = DataValidation(type="list", formula1='"Voluntaria,Obligatoria"', allow_blank=True)
    ws.add_data_validation(dv_tipo)

    for i in range(50):
        r = 5 + i
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center_align
            if c in (1, 2, 3, 4, 6):
                cell.fill = input_fill
                cell.font = data_font

        # Horas trabajadas: =IF(AND(C>0,D>0),(D-C)*24,"")
        ws.cell(row=r, column=5).value = f'=IF(AND(C{r}<>"",D{r}<>""),(D{r}-C{r})*24,"")'
        ws.cell(row=r, column=5).font = formula_font
        ws.cell(row=r, column=5).number_format = '0.00'

        # Horas extra: =IF(E>0, MAX(0, E-F), "")
        ws.cell(row=r, column=7).value = f'=IF(E{r}<>"",MAX(0,E{r}-F{r}),"")'
        ws.cell(row=r, column=7).font = formula_font
        ws.cell(row=r, column=7).number_format = '0.00'

        dv_tipo.add(ws.cell(row=r, column=8))
        ws.cell(row=r, column=8).fill = input_fill

    add_brand_footer(ws, 56, "H")

    # --- Sheet 3: Resumen Mensual ---
    ws2 = wb.create_sheet("Resumen Mensual")
    ws2.sheet_properties.tabColor = "BF360C"
    set_col_widths(ws2, [("A", 22), ("B", 16), ("C", 16), ("D", 16),
                         ("E", 18), ("F", 18), ("G", 18)])

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "Resumen Mensual de Horas Extras"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:G2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    # Config area
    ws2.cell(row=3, column=1, value="Tarifa hora base (€):").font = bold_font
    ws2.cell(row=3, column=2, value=12).font = data_font
    ws2.cell(row=3, column=2).fill = input_fill
    ws2.cell(row=3, column=2).border = thin_border

    headers2 = ["Empleado", "Total H. Extra", "H. Acumuladas Año", "Dentro Límite (≤80h)", "Coste ×1.75", "Coste ×2.0", "Coste Total Extra"]
    style_header_row(ws2, 5, headers2)

    for i in range(15):
        r = 6 + i
        ws2.cell(row=r, column=1).fill = input_fill
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=1).font = data_font

        ws2.cell(row=r, column=2).fill = input_fill
        ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=2).font = data_font

        ws2.cell(row=r, column=3).fill = input_fill
        ws2.cell(row=r, column=3).border = thin_border
        ws2.cell(row=r, column=3).font = data_font

        # Dentro límite: =IF(C<=80,"✓ OK","⚠ EXCEDE 80h")
        ws2.cell(row=r, column=4).value = f'=IF(C{r}="","",IF(C{r}<=80,"✓ OK","⚠ EXCEDE"))'
        ws2.cell(row=r, column=4).font = data_font
        ws2.cell(row=r, column=4).alignment = center_align
        ws2.cell(row=r, column=4).border = thin_border

        # Coste x1.75: =MIN(B, MAX(0, 80-C+B)) * tarifa * 1.75
        ws2.cell(row=r, column=5).value = f'=IF(B{r}="","",MIN(B{r},MAX(0,80-(C{r}-B{r})))*$B$3*1.75)'
        ws2.cell(row=r, column=5).font = formula_font
        ws2.cell(row=r, column=5).number_format = '€#,##0.00'
        ws2.cell(row=r, column=5).border = thin_border

        # Coste x2.0: =MAX(0, C-80) within this month's extras * tarifa * 2
        ws2.cell(row=r, column=6).value = f'=IF(B{r}="","",MAX(0,B{r}-MIN(B{r},MAX(0,80-(C{r}-B{r}))))*$B$3*2)'
        ws2.cell(row=r, column=6).font = formula_font
        ws2.cell(row=r, column=6).number_format = '€#,##0.00'
        ws2.cell(row=r, column=6).border = thin_border

        # Total = E + F
        ws2.cell(row=r, column=7).value = f'=IF(B{r}="","",E{r}+F{r})'
        ws2.cell(row=r, column=7).font = formula_font
        ws2.cell(row=r, column=7).number_format = '€#,##0.00'
        ws2.cell(row=r, column=7).border = thin_border

    # Totals
    r_tot = 22
    ws2.cell(row=r_tot, column=1, value="TOTALES").font = bold_font
    ws2.cell(row=r_tot, column=1).border = thin_border
    for c in [2, 5, 6, 7]:
        ws2.cell(row=r_tot, column=c).value = f"=SUM({get_column_letter(c)}6:{get_column_letter(c)}20)"
        ws2.cell(row=r_tot, column=c).font = formula_font
        ws2.cell(row=r_tot, column=c).border = thin_border
        if c >= 5:
            ws2.cell(row=r_tot, column=c).number_format = '€#,##0.00'

    add_brand_footer(ws2, 24, "G")

    fname = os.path.join(OUTPUT_DIR, "02-control-horas-extras.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# 03 — COSTE LABORAL MENSUAL
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Coste Laboral Mensual", [
        "Cómo usar esta plantilla:",
        "▸ Introduce los datos de nómina de cada empleado en la hoja 'Nóminas'.",
        "▸ La Seguridad Social empresa (~30%) se calcula automáticamente.",
        "▸ La hoja 'Ratio Coste Laboral' compara tu coste total con las ventas.",
        "▸ La hoja 'Previsión por Servicio' te ayuda a planificar personal por turno.",
        "",
        "Ratios de referencia por tipo de negocio:",
        "▸ Fast casual / comida rápida: 25-28%",
        "▸ Restaurante casual: 28-33%",
        "▸ Fine dining / alta cocina: 35-40%",
        "▸ Catering / eventos: 30-35%",
        "",
        "Semáforo automático:",
        "▸ 🟢 Verde: ratio < 30% (excelente control)",
        "▸ 🟡 Amarillo: 30-35% (vigilar)",
        "▸ 🔴 Rojo: > 35% (acción correctiva necesaria)",
    ])

    # --- Sheet 2: Nóminas ---
    ws = wb.create_sheet("Nóminas")
    ws.sheet_properties.tabColor = "2E7D32"
    set_col_widths(ws, [("A", 22), ("B", 18), ("C", 18), ("D", 16),
                        ("E", 18), ("F", 16)])

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Nóminas — Mes: _______________"
    ws["A1"].font = title_font
    ws.merge_cells("A2:F2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Empleado", "Puesto", "Salario Bruto Mes", "SS Empresa (30%)", "Coste Total", "Horas Contratadas"]
    style_header_row(ws, 4, headers)

    for i in range(20):
        r = 5 + i
        for c in [1, 2, 3, 6]:
            ws.cell(row=r, column=c).fill = input_fill
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = data_font
            if c == 3:
                ws.cell(row=r, column=c).number_format = '€#,##0.00'

        # SS = bruto * 0.30
        ws.cell(row=r, column=4).value = f'=IF(C{r}<>"",C{r}*0.30,"")'
        ws.cell(row=r, column=4).font = formula_font
        ws.cell(row=r, column=4).number_format = '€#,##0.00'
        ws.cell(row=r, column=4).border = thin_border

        # Coste total = bruto + SS
        ws.cell(row=r, column=5).value = f'=IF(C{r}<>"",C{r}+D{r},"")'
        ws.cell(row=r, column=5).font = formula_font
        ws.cell(row=r, column=5).number_format = '€#,##0.00'
        ws.cell(row=r, column=5).border = thin_border

    # Totals
    r_tot = 25
    ws.cell(row=r_tot, column=1, value="TOTALES").font = bold_font
    ws.cell(row=r_tot, column=1).border = thin_border
    for c in [3, 4, 5]:
        ws.cell(row=r_tot, column=c).value = f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}24)"
        ws.cell(row=r_tot, column=c).font = formula_font
        ws.cell(row=r_tot, column=c).number_format = '€#,##0.00'
        ws.cell(row=r_tot, column=c).border = thin_border

    add_brand_footer(ws, 27, "F")

    # --- Sheet 3: Ratio Coste Laboral ---
    ws2 = wb.create_sheet("Ratio Coste Laboral")
    ws2.sheet_properties.tabColor = "1B5E20"
    set_col_widths(ws2, [("A", 30), ("B", 22), ("C", 18)])

    ws2.merge_cells("A1:C1")
    ws2["A1"].value = "Ratio Coste Laboral vs Ventas"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:C2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    data_rows = [
        ("Ventas Netas del Mes (€):", None, True),
        ("Coste Laboral Total (€):", "=Nóminas!E25", False),
        ("", "", False),
        ("RATIO COSTE LABORAL (%):", "=IF(B4>0,B5/B4*100,0)", False),
        ("", "", False),
        ("Semáforo:", '=IF(B7<30,"🟢 EXCELENTE (<30%)",IF(B7<=35,"🟡 VIGILAR (30-35%)","🔴 ACCIÓN CORRECTIVA (>35%)"))', False),
    ]
    for idx, (label, val, is_input) in enumerate(data_rows):
        r = 4 + idx
        ws2.cell(row=r, column=1, value=label).font = bold_font
        ws2.cell(row=r, column=1).border = thin_border
        if val is not None:
            ws2.cell(row=r, column=2, value=val).font = formula_font
        if is_input:
            ws2.cell(row=r, column=2).fill = input_fill
        ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=2).number_format = '#,##0.00'

    # Reference table
    ws2.cell(row=12, column=1, value="Ratios de Referencia por Tipo").font = section_font
    ref_headers = ["Tipo de Negocio", "Ratio Objetivo", "Rango Aceptable"]
    style_header_row(ws2, 13, ref_headers)
    ref_data = [
        ("Fast Casual / Comida Rápida", "25-28%", "22-30%"),
        ("Restaurante Casual", "28-33%", "25-35%"),
        ("Fine Dining / Alta Cocina", "35-40%", "33-42%"),
        ("Catering / Eventos", "30-35%", "28-38%"),
        ("Cafetería / Brunch", "28-32%", "25-35%"),
        ("Bar / Cocktails", "25-30%", "22-33%"),
    ]
    for idx, (tipo, obj, rango) in enumerate(ref_data):
        r = 14 + idx
        ws2.cell(row=r, column=1, value=tipo).font = data_font
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=obj).font = data_font
        ws2.cell(row=r, column=2).alignment = center_align
        ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=3, value=rango).font = data_font
        ws2.cell(row=r, column=3).alignment = center_align
        ws2.cell(row=r, column=3).border = thin_border

    add_brand_footer(ws2, 22, "C")

    # --- Sheet 4: Previsión por Servicio ---
    ws3 = wb.create_sheet("Previsión por Servicio")
    ws3.sheet_properties.tabColor = "4CAF50"
    set_col_widths(ws3, [("A", 30), ("B", 18), ("C", 18), ("D", 18)])

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "Previsión de Personal por Servicio"
    ws3["A1"].font = title_font
    ws3.merge_cells("A2:D2")
    ws3["A2"].value = BRAND_LINE
    ws3["A2"].font = subtitle_font

    # Inputs
    inputs = [
        ("Covers previstos / día:", 4),
        ("Ratio covers/cocinero:", 5),
        ("Ratio covers/camarero:", 6),
        ("Ratio covers/barman:", 7),
    ]
    for label, r in inputs:
        ws3.cell(row=r, column=1, value=label).font = bold_font
        ws3.cell(row=r, column=1).border = thin_border
        ws3.cell(row=r, column=2).fill = input_fill
        ws3.cell(row=r, column=2).border = thin_border

    # Default values
    ws3.cell(row=4, column=2, value=80).font = data_font
    ws3.cell(row=5, column=2, value=20).font = data_font
    ws3.cell(row=6, column=2, value=12).font = data_font
    ws3.cell(row=7, column=2, value=25).font = data_font

    # Results
    ws3.cell(row=9, column=1, value="PERSONAL NECESARIO POR TURNO").font = section_font
    results = [
        ("Cocineros necesarios:", "=CEILING(B4/B5,1)"),
        ("Camareros necesarios:", "=CEILING(B4/B6,1)"),
        ("Barmans necesarios:", "=CEILING(B4/B7,1)"),
        ("TOTAL personal/turno:", "=B10+B11+B12"),
    ]
    for idx, (label, formula) in enumerate(results):
        r = 10 + idx
        ws3.cell(row=r, column=1, value=label).font = bold_font
        ws3.cell(row=r, column=1).border = thin_border
        ws3.cell(row=r, column=2, value=formula).font = formula_font
        ws3.cell(row=r, column=2).alignment = center_align
        ws3.cell(row=r, column=2).border = thin_border

    # Coste estimado
    ws3.cell(row=15, column=1, value="Salario medio mensual (€):").font = bold_font
    ws3.cell(row=15, column=2, value=1800).font = data_font
    ws3.cell(row=15, column=2).fill = input_fill
    ws3.cell(row=15, column=2).border = thin_border

    ws3.cell(row=16, column=1, value="Coste laboral estimado/mes (€):").font = bold_font
    ws3.cell(row=16, column=2, value="=B13*B15*1.30").font = formula_font
    ws3.cell(row=16, column=2).number_format = '€#,##0.00'
    ws3.cell(row=16, column=2).border = thin_border

    add_brand_footer(ws3, 19, "D")

    fname = os.path.join(OUTPUT_DIR, "03-coste-laboral-mensual.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# 04 — ONBOARDING NUEVO EMPLEADO
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Onboarding Nuevo Empleado", [
        "Cómo usar esta plantilla:",
        "▸ Abre un checklist nuevo para cada incorporación.",
        "▸ Asigna un responsable a cada tarea y establece fecha límite.",
        "▸ Marca ✓ cuando se complete cada paso.",
        "▸ Revisa el checklist completo antes de finalizar el periodo de prueba.",
        "",
        "Categorías:",
        "▸ Azul = Documentación legal y administrativa",
        "▸ Naranja = Formación obligatoria",
        "▸ Verde = Equipamiento y accesos",
        "▸ Morado = Formación operativa",
        "▸ Rojo = Periodo de prueba y evaluación",
    ])

    ws = wb.create_sheet("Checklist Onboarding")
    ws.sheet_properties.tabColor = "FF6F00"
    set_col_widths(ws, [("A", 5), ("B", 45), ("C", 16), ("D", 18),
                        ("E", 16), ("F", 12), ("G", 20)])

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Checklist Onboarding — Empleado: _______________"
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    # Employee data
    ws.cell(row=3, column=1, value="Puesto:").font = bold_font
    ws.cell(row=3, column=2).fill = input_fill
    ws.cell(row=3, column=2).border = thin_border
    ws.cell(row=3, column=4, value="Fecha alta:").font = bold_font
    ws.cell(row=3, column=5).fill = input_fill
    ws.cell(row=3, column=5).border = thin_border

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    ws.add_data_validation(dv)

    sections = [
        ("📋 DOCUMENTACIÓN LEGAL Y ADMINISTRATIVA", DOC_COLOR, [
            "Firma de contrato de trabajo",
            "Copia DNI / NIE / Pasaporte",
            "Datos cuenta bancaria para nómina",
            "Alta en Seguridad Social (TA.2/S)",
            "Parte de alta PRL — reconocimiento médico programado",
            "Firma cláusula protección de datos (RGPD)",
            "Firma compromiso confidencialidad",
            "Entrega copia convenio colectivo aplicable",
            "Registro jornada — explicar sistema de fichaje",
            "Foto para ficha de personal",
        ]),
        ("🎓 FORMACIÓN OBLIGATORIA", FORMACION_COLOR, [
            "Carnet manipulador de alimentos vigente",
            "Formación APPCC — protocolos del establecimiento",
            "Formación alérgenos — los 14 alérgenos de declaración obligatoria",
            "PRL específica del puesto — riesgos y medidas preventivas",
            "Plan de emergencia y evacuación — rutas y puntos de encuentro",
            "Protocolo acoso laboral — información y canal de denuncia",
            "Normativa de uniformidad e higiene personal",
            "Formación en igualdad (si aplica según plantilla)",
        ]),
        ("👕 EQUIPAMIENTO Y ACCESOS", EQUIPO_COLOR, [
            "Uniforme completo entregado (tallas verificadas)",
            "Calzado de seguridad antideslizante",
            "Taquilla asignada con llave/candado",
            "Llaves / tarjeta de acceso al local",
            "Código alarma comunicado",
            "Acceso sistema TPV / caja registradora",
            "Acceso herramientas digitales (email, app turnos, etc.)",
            "Placa identificativa / nombre",
            "EPI adicional si aplica (guantes, delantal térmico, etc.)",
        ]),
        ("🏢 FORMACIÓN OPERATIVA", OPERATIVA_COLOR, [
            "Tour completo de instalaciones",
            "Presentación al equipo y responsables directos",
            "Manual del puesto entregado y explicado",
            "Protocolo de limpieza y desinfección",
            "Sistema de caja / TPV — práctica supervisada",
            "Carta / menú — estudio y degustación",
            "Protocolo de atención al cliente",
            "Sistema de reservas (si aplica)",
            "Protocolo de aperturas y cierres",
            "Gestión de residuos y reciclaje",
            "Sistema de pedidos a proveedores (si aplica)",
            "Protocolo de quejas y reclamaciones",
        ]),
        ("📊 PERIODO DE PRUEBA Y EVALUACIÓN", PRUEBA_COLOR, [
            "Objetivos semana 1 definidos y comunicados",
            "Objetivos semana 2 definidos y comunicados",
            "Objetivos semana 4 (primer mes) definidos",
            "Evaluación informal día 3 — primeras impresiones",
            "Evaluación formal día 15 — revisión intermedia",
            "Evaluación formal día 30 — decisión continuidad",
            "Feedback del equipo recogido",
            "Plan de desarrollo si se confirma en puesto",
        ]),
    ]

    row = 5
    for section_title, color, tasks in sections:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        col_headers = ["#", "Tarea", "Categoría", "Responsable", "Fecha Límite", "✓", "Notas"]
        style_header_row(ws, row, col_headers)
        row += 1

        cat_name = section_title.split(" ", 1)[1] if " " in section_title else section_title
        for idx, task in enumerate(tasks, 1):
            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=cat_name[:14]).font = data_font
            ws.cell(row=row, column=3).fill = fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border

            ws.cell(row=row, column=5).fill = input_fill
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=5).number_format = 'DD/MM/YYYY'

            check = ws.cell(row=row, column=6)
            check.font = checkbox_font
            check.alignment = center_align
            check.border = thin_border
            dv.add(check)

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    # Summary
    ws.cell(row=row, column=1, value="RESUMEN").font = section_font
    row += 1
    ws.cell(row=row, column=2, value="Total tareas completadas:").font = bold_font
    ws.cell(row=row, column=3).value = f'=COUNTIF(F6:F{row-2},"✓")'
    ws.cell(row=row, column=3).font = formula_font
    ws.cell(row=row, column=3).border = thin_border
    row += 1
    ws.cell(row=row, column=2, value="Total tareas:").font = bold_font
    ws.cell(row=row, column=3, value=47).font = data_font
    ws.cell(row=row, column=3).border = thin_border
    row += 1
    ws.cell(row=row, column=2, value="Progreso (%):").font = bold_font
    ws.cell(row=row, column=3).value = f'=IF({row-2}>0,{row-2}/{row-1}*100,0)'
    ws.cell(row=row, column=3).font = formula_font
    ws.cell(row=row, column=3).number_format = '0.0"%"'
    ws.cell(row=row, column=3).border = thin_border

    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    add_brand_footer(ws, row, "G")

    fname = os.path.join(OUTPUT_DIR, "04-onboarding-nuevo-empleado.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# 05 — PLANIFICACIÓN DE VACACIONES
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Planificación de Vacaciones Anual", [
        "Cómo usar esta plantilla:",
        "▸ En el Calendario Anual, marca con código de color los días de cada empleado.",
        "▸ Azul = Vacaciones, Rojo = Baja médica, Verde = Festivo, Naranja = Permiso.",
        "▸ En 'Solicitudes' registra cada petición y su estado.",
        "▸ Los días usados y restantes se calculan automáticamente (de 30 días legales).",
        "",
        "Legislación española:",
        "▸ 30 días naturales de vacaciones al año (Art. 38 ET).",
        "▸ O 22 días laborables según convenio habitual hostelería.",
        "▸ Al menos 2 semanas consecutivas.",
        "▸ Fijar calendario vacaciones con 2 meses de antelación.",
        "",
        "Hoja Cobertura:",
        "▸ Define mínimo de personal por turno y quién cubre a quién.",
    ])

    # --- Sheet 2: Calendario Anual ---
    ws = wb.create_sheet("Calendario Anual")
    ws.sheet_properties.tabColor = "1565C0"

    ws.column_dimensions["A"].width = 20
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    for i, m in enumerate(months):
        ws.column_dimensions[get_column_letter(i + 2)].width = 8

    ws.merge_cells("A1:M1")
    ws["A1"].value = "Calendario de Vacaciones — Año: ______"
    ws["A1"].font = title_font
    ws.merge_cells("A2:M2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    # Legend
    legend = [("V=Vacaciones", VACACIONES_COLOR), ("B=Baja", BAJA_COLOR),
              ("F=Festivo", FESTIVO_COLOR), ("P=Permiso", PERMISO_COLOR)]
    for ci, (txt, color) in enumerate(legend):
        cell = ws.cell(row=3, column=2 + ci, value=txt)
        cell.font = Font(name="Calibri", size=9, bold=True)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    headers = ["Empleado"] + months + ["Días Usados"]
    style_header_row(ws, 5, headers)
    ws.column_dimensions[get_column_letter(14)].width = 14

    dv_vac = DataValidation(type="list", formula1='"V,B,F,P,"', allow_blank=True)
    ws.add_data_validation(dv_vac)

    for i in range(15):
        r = 6 + i
        ws.cell(row=r, column=1).fill = input_fill
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=1).font = data_font

        for c in range(2, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = input_fill
            dv_vac.add(cell)

        # Días usados: count "V" entries
        parts = [f'IF({get_column_letter(c)}{r}="V",1,0)' for c in range(2, 14)]
        ws.cell(row=r, column=14).value = "=" + "+".join(parts)
        ws.cell(row=r, column=14).font = formula_font
        ws.cell(row=r, column=14).alignment = center_align
        ws.cell(row=r, column=14).border = thin_border

    add_brand_footer(ws, 23, "N")

    # --- Sheet 3: Solicitudes ---
    ws2 = wb.create_sheet("Solicitudes")
    ws2.sheet_properties.tabColor = "0D47A1"
    set_col_widths(ws2, [("A", 22), ("B", 16), ("C", 16), ("D", 14),
                         ("E", 14), ("F", 16), ("G", 18)])

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "Solicitudes de Vacaciones"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:G2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    headers2 = ["Empleado", "Fecha Inicio", "Fecha Fin", "Días Solicit.", "Días Usados YTD", "Días Restantes", "Estado"]
    style_header_row(ws2, 4, headers2)

    dv_estado = DataValidation(type="list", formula1='"Pendiente,Aprobado,Denegado"', allow_blank=True)
    ws2.add_data_validation(dv_estado)

    for i in range(30):
        r = 5 + i
        for c in [1, 2, 3]:
            ws2.cell(row=r, column=c).fill = input_fill
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).font = data_font
        for c in [2, 3]:
            ws2.cell(row=r, column=c).number_format = 'DD/MM/YYYY'

        # Días solicitados: fecha fin - fecha inicio + 1
        ws2.cell(row=r, column=4).value = f'=IF(AND(B{r}<>"",C{r}<>""),C{r}-B{r}+1,"")'
        ws2.cell(row=r, column=4).font = formula_font
        ws2.cell(row=r, column=4).alignment = center_align
        ws2.cell(row=r, column=4).border = thin_border

        # Días usados YTD — manual input
        ws2.cell(row=r, column=5).fill = input_fill
        ws2.cell(row=r, column=5).border = thin_border
        ws2.cell(row=r, column=5).alignment = center_align

        # Días restantes = 30 - usados
        ws2.cell(row=r, column=6).value = f'=IF(E{r}<>"",30-E{r},"")'
        ws2.cell(row=r, column=6).font = formula_font
        ws2.cell(row=r, column=6).alignment = center_align
        ws2.cell(row=r, column=6).border = thin_border

        estado_cell = ws2.cell(row=r, column=7)
        estado_cell.fill = input_fill
        estado_cell.border = thin_border
        estado_cell.alignment = center_align
        dv_estado.add(estado_cell)

    add_brand_footer(ws2, 36, "G")

    # --- Sheet 4: Cobertura ---
    ws3 = wb.create_sheet("Cobertura")
    ws3.sheet_properties.tabColor = "4CAF50"
    set_col_widths(ws3, [("A", 22), ("B", 16), ("C", 22), ("D", 22), ("E", 20)])

    ws3.merge_cells("A1:E1")
    ws3["A1"].value = "Plan de Cobertura durante Vacaciones"
    ws3["A1"].font = title_font
    ws3.merge_cells("A2:E2")
    ws3["A2"].value = BRAND_LINE
    ws3["A2"].font = subtitle_font

    # Minimum staff
    ws3.cell(row=4, column=1, value="PERSONAL MÍNIMO POR TURNO").font = section_font
    min_headers = ["Turno", "Mínimo Personal", "Cocina", "Sala", "Barra"]
    style_header_row(ws3, 5, min_headers)
    turnos = [("Mañana", ""), ("Tarde", ""), ("Noche", "")]
    for idx, (turno, _) in enumerate(turnos):
        r = 6 + idx
        ws3.cell(row=r, column=1, value=turno).font = data_font
        ws3.cell(row=r, column=1).border = thin_border
        for c in range(2, 6):
            ws3.cell(row=r, column=c).fill = input_fill
            ws3.cell(row=r, column=c).border = thin_border
            ws3.cell(row=r, column=c).alignment = center_align

    # Coverage pairs
    ws3.cell(row=10, column=1, value="COBERTURA: QUIÉN SUSTITUYE A QUIÉN").font = section_font
    cov_headers = ["Empleado Ausente", "Puesto", "Sustituto 1", "Sustituto 2", "Notas"]
    style_header_row(ws3, 11, cov_headers)
    for i in range(15):
        r = 12 + i
        for c in range(1, 6):
            ws3.cell(row=r, column=c).fill = input_fill
            ws3.cell(row=r, column=c).border = thin_border
            ws3.cell(row=r, column=c).font = data_font

    add_brand_footer(ws3, 29, "E")

    fname = os.path.join(OUTPUT_DIR, "05-planificacion-vacaciones.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# 06 — EVALUACIÓN DE DESEMPEÑO
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Evaluación de Desempeño", [
        "Cómo usar esta plantilla:",
        "▸ Completa una ficha por empleado y periodo de evaluación.",
        "▸ Puntúa cada competencia de 1 (deficiente) a 5 (excelente).",
        "▸ La media se calcula automáticamente.",
        "▸ Usa la hoja 'Histórico' para seguir la evolución trimestral.",
        "",
        "Escala de puntuación:",
        "▸ 1 = Deficiente — No cumple requisitos mínimos",
        "▸ 2 = Mejorable — Cumple parcialmente, necesita supervisión constante",
        "▸ 3 = Adecuado — Cumple lo esperado para el puesto",
        "▸ 4 = Bueno — Supera expectativas en varias áreas",
        "▸ 5 = Excelente — Referencia para el equipo",
        "",
        "Frecuencia recomendada:",
        "▸ Evaluación formal cada 3 meses (trimestral).",
        "▸ Feedback informal continuo — al menos 1 conversación/semana.",
    ])

    # --- Sheet 2: Ficha Evaluación ---
    ws = wb.create_sheet("Ficha Evaluación")
    ws.sheet_properties.tabColor = "7B1FA2"
    set_col_widths(ws, [("A", 5), ("B", 35), ("C", 14), ("D", 40)])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Ficha de Evaluación de Desempeño"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    # Employee info
    info_fields = [
        ("Empleado:", 4), ("Puesto:", 5), ("Periodo evaluado:", 6),
        ("Evaluador:", 7), ("Fecha evaluación:", 8),
    ]
    for label, r in info_fields:
        ws.cell(row=r, column=2, value=label).font = bold_font
        ws.cell(row=r, column=3).fill = input_fill
        ws.cell(row=r, column=3).border = thin_border
        ws.merge_cells(f"C{r}:D{r}")

    # Competencies
    ws.cell(row=10, column=1, value="").font = section_font
    ws.merge_cells("A10:D10")
    ws["A10"].value = "COMPETENCIAS (puntúa 1-5)"
    ws["A10"].font = section_font

    comp_headers = ["#", "Competencia", "Puntuación (1-5)", "Observaciones"]
    style_header_row(ws, 11, comp_headers)

    competencies = [
        "Puntualidad y asistencia",
        "Presentación personal e higiene",
        "Trabajo en equipo y colaboración",
        "Iniciativa y proactividad",
        "Limpieza y orden del puesto",
        "Rapidez y eficiencia",
        "Atención al cliente",
        "Conocimiento de la carta / producto",
        "Gestión del estrés en servicio",
        "Comunicación con compañeros y superiores",
    ]

    dv_score = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    dv_score.error = "Introduce un valor entre 1 y 5"
    dv_score.prompt = "1=Deficiente, 2=Mejorable, 3=Adecuado, 4=Bueno, 5=Excelente"
    ws.add_data_validation(dv_score)

    for idx, comp in enumerate(competencies):
        r = 12 + idx
        ws.cell(row=r, column=1, value=idx + 1).font = data_font
        ws.cell(row=r, column=1).alignment = center_align
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2, value=comp).font = data_font
        ws.cell(row=r, column=2).alignment = left_align
        ws.cell(row=r, column=2).border = thin_border

        score_cell = ws.cell(row=r, column=3)
        score_cell.fill = input_fill
        score_cell.alignment = center_align
        score_cell.border = thin_border
        score_cell.font = data_font
        dv_score.add(score_cell)

        ws.cell(row=r, column=4).fill = input_fill
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).alignment = left_align

    # Average
    r_avg = 22
    ws.cell(row=r_avg, column=1).border = thin_border
    ws.cell(row=r_avg, column=2, value="PUNTUACIÓN MEDIA").font = bold_font
    ws.cell(row=r_avg, column=2).border = thin_border
    ws.cell(row=r_avg, column=3).value = "=AVERAGE(C12:C21)"
    ws.cell(row=r_avg, column=3).font = formula_font
    ws.cell(row=r_avg, column=3).alignment = center_align
    ws.cell(row=r_avg, column=3).border = thin_border
    ws.cell(row=r_avg, column=3).number_format = '0.0'

    r_level = 23
    ws.cell(row=r_level, column=2, value="NIVEL").font = bold_font
    ws.cell(row=r_level, column=2).border = thin_border
    ws.cell(row=r_level, column=3).value = '=IF(C22>=4.5,"⭐ EXCELENTE",IF(C22>=3.5,"✓ BUENO",IF(C22>=2.5,"→ ADECUADO",IF(C22>=1.5,"⚠ MEJORABLE","✗ DEFICIENTE"))))'
    ws.cell(row=r_level, column=3).font = formula_font
    ws.cell(row=r_level, column=3).border = thin_border

    # Qualitative sections
    qual_sections = [
        (25, "FORTALEZAS DESTACADAS"),
        (30, "ÁREAS DE MEJORA"),
        (35, "OBJETIVOS PARA PRÓXIMO PERIODO"),
    ]
    for start_r, title in qual_sections:
        ws.merge_cells(f"A{start_r}:D{start_r}")
        ws[f"A{start_r}"].value = title
        ws[f"A{start_r}"].font = section_font
        for r in range(start_r + 1, start_r + 4):
            ws.merge_cells(f"B{r}:D{r}")
            ws.cell(row=r, column=2).fill = input_fill
            ws.cell(row=r, column=2).border = thin_border

    # Signatures
    ws.merge_cells("A40:D40")
    ws["A40"].value = "Firma evaluador: _________________     Firma empleado: _________________"
    ws["A40"].font = small_font
    ws.merge_cells("A41:D41")
    ws["A41"].value = "Fecha: ___/___/______"
    ws["A41"].font = small_font
    add_brand_footer(ws, 43, "D")

    # --- Sheet 3: Histórico ---
    ws2 = wb.create_sheet("Histórico")
    ws2.sheet_properties.tabColor = "4A148C"
    set_col_widths(ws2, [("A", 22), ("B", 14), ("C", 14), ("D", 14),
                         ("E", 14), ("F", 14), ("G", 16)])

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "Histórico de Evaluaciones"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:G2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    hist_headers = ["Empleado", "Q1", "Q2", "Q3", "Q4", "Media Anual", "Tendencia"]
    style_header_row(ws2, 4, hist_headers)

    for i in range(15):
        r = 5 + i
        ws2.cell(row=r, column=1).fill = input_fill
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=1).font = data_font

        for c in range(2, 6):
            ws2.cell(row=r, column=c).fill = input_fill
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).alignment = center_align

        # Media anual
        ws2.cell(row=r, column=6).value = f'=IF(COUNT(B{r}:E{r})>0,AVERAGE(B{r}:E{r}),"")'
        ws2.cell(row=r, column=6).font = formula_font
        ws2.cell(row=r, column=6).alignment = center_align
        ws2.cell(row=r, column=6).border = thin_border
        ws2.cell(row=r, column=6).number_format = '0.0'

        # Tendencia: comparing last two available
        ws2.cell(row=r, column=7).value = f'=IF(AND(E{r}<>"",D{r}<>""),IF(E{r}>D{r},"↑ Mejora",IF(E{r}<D{r},"↓ Baja","→ Estable")),"")'
        ws2.cell(row=r, column=7).font = data_font
        ws2.cell(row=r, column=7).alignment = center_align
        ws2.cell(row=r, column=7).border = thin_border

    add_brand_footer(ws2, 22, "G")

    fname = os.path.join(OUTPUT_DIR, "06-evaluacion-desempeno.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# 07 — DIRECTORIO DE PLANTILLA
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Directorio de Plantilla", [
        "Cómo usar esta plantilla:",
        "▸ Registra todos los datos de cada empleado en la hoja 'Plantilla'.",
        "▸ Usa filtros de Excel para buscar por puesto, tipo de contrato, etc.",
        "▸ La hoja 'Vencimientos' te alerta de contratos a punto de expirar.",
        "",
        "Datos incluidos:",
        "▸ Datos personales: nombre, DNI/NIE, teléfono, email, contacto emergencia.",
        "▸ Datos laborales: puesto, contrato, fecha alta, jornada, salario.",
        "▸ Datos operativos: talla uniforme, taquilla, alérgenos propios.",
        "",
        "Protección de datos (RGPD):",
        "▸ Este archivo contiene datos personales — proteger con contraseña.",
        "▸ Solo acceso para RRHH y dirección.",
        "▸ No compartir por email sin cifrar.",
    ])

    # --- Sheet 2: Plantilla ---
    ws = wb.create_sheet("Plantilla")
    ws.sheet_properties.tabColor = "00695C"

    columns = [
        ("A", "Nombre Completo", 24),
        ("B", "DNI / NIE", 16),
        ("C", "Puesto", 18),
        ("D", "Tipo Contrato", 16),
        ("E", "Fecha Alta", 14),
        ("F", "Fecha Fin Contrato", 16),
        ("G", "Jornada", 14),
        ("H", "Horas Semanales", 14),
        ("I", "Salario Bruto Anual", 18),
        ("J", "Teléfono", 14),
        ("K", "Email", 22),
        ("L", "Contacto Emergencia", 20),
        ("M", "Alérgenos Propios", 18),
        ("N", "Talla Uniforme", 14),
        ("O", "Taquilla", 10),
    ]
    for letter, _, width in columns:
        ws.column_dimensions[letter].width = width

    ws.merge_cells("A1:O1")
    ws["A1"].value = "Directorio de Plantilla"
    ws["A1"].font = title_font
    ws.merge_cells("A2:O2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = [name for _, name, _ in columns]
    style_header_row(ws, 4, headers)

    dv_contrato = DataValidation(type="list", formula1='"Indefinido,Temporal,Prácticas,Formación,Fijo-discontinuo"', allow_blank=True)
    dv_jornada = DataValidation(type="list", formula1='"Completa,Parcial"', allow_blank=True)
    ws.add_data_validation(dv_contrato)
    ws.add_data_validation(dv_jornada)

    for i in range(30):
        r = 5 + i
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = thin_border
            cell.font = data_font
            cell.alignment = left_align if c in (1, 3, 11, 12, 13) else center_align

        ws.cell(row=r, column=5).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=6).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=9).number_format = '€#,##0.00'
        dv_contrato.add(ws.cell(row=r, column=4))
        dv_jornada.add(ws.cell(row=r, column=7))

    ws.auto_filter.ref = "A4:O34"
    add_brand_footer(ws, 36, "O")

    # --- Sheet 3: Vencimientos ---
    ws2 = wb.create_sheet("Vencimientos")
    ws2.sheet_properties.tabColor = "004D40"
    set_col_widths(ws2, [("A", 24), ("B", 18), ("C", 18), ("D", 18), ("E", 18)])

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Alertas de Vencimientos"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:E2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    ws2.cell(row=3, column=1, value="Fecha de hoy:").font = bold_font
    ws2.cell(row=3, column=2, value="=TODAY()").font = formula_font
    ws2.cell(row=3, column=2).number_format = 'DD/MM/YYYY'
    ws2.cell(row=3, column=2).border = thin_border

    ws2.cell(row=5, column=1, value="CONTRATOS A EXPIRAR").font = section_font
    venc_headers = ["Empleado", "Tipo Contrato", "Fecha Fin", "Días Restantes", "Alerta"]
    style_header_row(ws2, 6, venc_headers)

    for i in range(15):
        r = 7 + i
        # Reference from Plantilla sheet
        ws2.cell(row=r, column=1).value = f"=IF(Plantilla!A{5+i}<>\"\",Plantilla!A{5+i},\"\")"
        ws2.cell(row=r, column=1).font = data_font
        ws2.cell(row=r, column=1).border = thin_border

        ws2.cell(row=r, column=2).value = f"=IF(Plantilla!D{5+i}<>\"\",Plantilla!D{5+i},\"\")"
        ws2.cell(row=r, column=2).font = data_font
        ws2.cell(row=r, column=2).border = thin_border

        ws2.cell(row=r, column=3).value = f"=IF(Plantilla!F{5+i}<>\"\",Plantilla!F{5+i},\"\")"
        ws2.cell(row=r, column=3).font = data_font
        ws2.cell(row=r, column=3).number_format = 'DD/MM/YYYY'
        ws2.cell(row=r, column=3).border = thin_border

        # Días restantes
        ws2.cell(row=r, column=4).value = f'=IF(C{r}<>"",C{r}-$B$3,"")'
        ws2.cell(row=r, column=4).font = formula_font
        ws2.cell(row=r, column=4).alignment = center_align
        ws2.cell(row=r, column=4).border = thin_border

        # Alerta
        ws2.cell(row=r, column=5).value = f'=IF(D{r}="","",IF(D{r}<=0,"❌ VENCIDO",IF(D{r}<=30,"🔴 <30 días",IF(D{r}<=60,"🟡 <60 días",IF(D{r}<=90,"🟢 <90 días","OK")))))'
        ws2.cell(row=r, column=5).font = data_font
        ws2.cell(row=r, column=5).alignment = center_align
        ws2.cell(row=r, column=5).border = thin_border

    # Carnets section
    ws2.cell(row=24, column=1, value="CARNETS Y CERTIFICADOS A RENOVAR").font = section_font
    cert_headers = ["Empleado", "Certificado", "Fecha Caducidad", "Días Restantes", "Alerta"]
    style_header_row(ws2, 25, cert_headers)
    for i in range(10):
        r = 26 + i
        for c in range(1, 4):
            ws2.cell(row=r, column=c).fill = input_fill
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).font = data_font
        ws2.cell(row=r, column=3).number_format = 'DD/MM/YYYY'

        ws2.cell(row=r, column=4).value = f'=IF(C{r}<>"",C{r}-$B$3,"")'
        ws2.cell(row=r, column=4).font = formula_font
        ws2.cell(row=r, column=4).alignment = center_align
        ws2.cell(row=r, column=4).border = thin_border

        ws2.cell(row=r, column=5).value = f'=IF(D{r}="","",IF(D{r}<=0,"❌ VENCIDO",IF(D{r}<=30,"🔴 URGENTE",IF(D{r}<=60,"🟡 PRONTO","✓ OK"))))'
        ws2.cell(row=r, column=5).font = data_font
        ws2.cell(row=r, column=5).alignment = center_align
        ws2.cell(row=r, column=5).border = thin_border

    add_brand_footer(ws2, 38, "E")

    fname = os.path.join(OUTPUT_DIR, "07-directorio-plantilla.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# BONUS-01 — BRIEFING CAMBIO DE TURNO
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS · Briefing de Cambio de Turno", [
        "Cómo usar esta plantilla:",
        "▸ El turno saliente completa este briefing ANTES de irse.",
        "▸ El turno entrante lo revisa y firma al llegar.",
        "▸ Imprime una copia diaria o usa en tablet.",
        "▸ Archiva los briefings para trazabilidad.",
        "",
        "Secciones incluidas:",
        "▸ Reservas pendientes y VIPs del siguiente turno.",
        "▸ Incidencias: averías, falta de stock, quejas.",
        "▸ Tareas pendientes que no se completaron.",
        "▸ Stock bajo que necesita pedido urgente.",
        "▸ Personal: ausencias y cambios de último momento.",
    ])

    ws = wb.create_sheet("Briefing")
    ws.sheet_properties.tabColor = "F57F17"
    set_col_widths(ws, [("A", 5), ("B", 50), ("C", 20), ("D", 20)])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Briefing de Cambio de Turno"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    # Header info
    info = [
        ("Fecha:", 4, 2), ("Turno Saliente:", 4, 3), ("Turno Entrante:", 4, 4),
    ]
    for label, r, c in info:
        ws.cell(row=r, column=c, value=label).font = bold_font
    ws.cell(row=5, column=2).fill = input_fill
    ws.cell(row=5, column=2).border = thin_border
    ws.cell(row=5, column=3).fill = input_fill
    ws.cell(row=5, column=3).border = thin_border
    ws.cell(row=5, column=4).fill = input_fill
    ws.cell(row=5, column=4).border = thin_border

    sections = [
        ("🍽️ RESERVAS PENDIENTES / VIPs", [
            ("Hora", "Mesa/Zona", "Nombre", "Notas (VIP, alergia, celebración...)"),
        ], 5),
        ("⚠️ INCIDENCIAS DEL TURNO", [
            ("Hora", "Incidencia", "Gravedad", "Acción tomada"),
        ], 5),
        ("📋 TAREAS PENDIENTES PARA SIGUIENTE TURNO", [
            ("Prioridad", "Tarea", "Responsable", "Notas"),
        ], 5),
        ("📦 STOCK BAJO — PEDIDOS URGENTES", [
            ("Producto", "Cantidad Actual", "Cantidad Necesaria", "Proveedor"),
        ], 5),
        ("👥 PERSONAL: AUSENCIAS Y CAMBIOS", [
            ("Empleado", "Situación", "Sustituto", "Notas"),
        ], 4),
    ]

    row = 7
    for section_title, header_sets, num_rows in sections:
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        for hdrs in header_sets:
            for ci, h in enumerate(hdrs):
                cell = ws.cell(row=row, column=ci + 1, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

            for _ in range(num_rows):
                for c in range(1, len(hdrs) + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = input_fill
                    cell.border = thin_border
                    cell.font = data_font
                row += 1
        row += 1

    # Signatures
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "OBSERVACIONES GENERALES:"
    ws[f"A{row}"].font = section_font
    row += 1
    for _ in range(3):
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"].fill = input_fill
        ws[f"A{row}"].border = thin_border
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "Firma turno saliente: _________________     Firma turno entrante: _________________"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "Hora entrega: ______     Hora recepción: ______"
    ws[f"A{row}"].font = small_font
    row += 1
    add_brand_footer(ws, row, "D")

    fname = os.path.join(OUTPUT_DIR, "BONUS-01-briefing-cambio-turno.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# BONUS-02 — CALCULADORA PLANTILLA ÓPTIMA
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS · Calculadora de Plantilla Óptima", [
        "Cómo usar esta plantilla:",
        "▸ Introduce los datos de tu negocio en las celdas verdes.",
        "▸ Las fórmulas calculan el personal óptimo automáticamente.",
        "▸ Compara con tu plantilla actual para detectar sobre/infra-dimensionamiento.",
        "▸ Consulta los ratios de referencia en la hoja 'Ratios por Tipo'.",
        "",
        "Cálculos incluidos:",
        "▸ Cocineros, camareros y barmans necesarios por turno.",
        "▸ Coste laboral estimado mensual.",
        "▸ Comparativa automática con plantilla actual.",
        "",
        "Ratios de referencia (covers por empleado):",
        "▸ Casual: 1 cocinero/20 covers, 1 camarero/12 covers.",
        "▸ Fine dining: 1 cocinero/8 covers, 1 camarero/6 covers.",
        "▸ Fast casual: 1 cocinero/30 covers, 1 camarero/20 covers.",
    ])

    # --- Sheet 2: Calculadora ---
    ws = wb.create_sheet("Calculadora")
    ws.sheet_properties.tabColor = "E65100"
    set_col_widths(ws, [("A", 35), ("B", 18), ("C", 18), ("D", 22)])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Calculadora de Plantilla Óptima"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    # Input section
    ws.cell(row=4, column=1, value="DATOS DE TU NEGOCIO").font = section_font

    inputs = [
        ("Covers / día promedio:", 5, 80),
        ("Días apertura / semana:", 6, 6),
        ("Servicios (1=almuerzo, 2=cena, 3=ambos):", 7, 3),
        ("Tipo negocio (1=casual, 2=fine, 3=fast):", 8, 1),
        ("Salario medio bruto mensual (€):", 9, 1800),
    ]
    for label, r, default in inputs:
        ws.cell(row=r, column=1, value=label).font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=default).font = data_font
        ws.cell(row=r, column=2).fill = input_fill
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=2).alignment = center_align

    # Ratio lookup (based on type)
    ws.cell(row=11, column=1, value="RATIOS APLICADOS (auto según tipo)").font = section_font

    ratio_rows = [
        ("Covers / cocinero:", 12, '=IF(B8=1,20,IF(B8=2,8,30))'),
        ("Covers / camarero:", 13, '=IF(B8=1,12,IF(B8=2,6,20))'),
        ("Covers / barman:", 14, '=IF(B8=1,25,IF(B8=2,15,35))'),
    ]
    for label, r, formula in ratio_rows:
        ws.cell(row=r, column=1, value=label).font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=formula).font = formula_font
        ws.cell(row=r, column=2).alignment = center_align
        ws.cell(row=r, column=2).border = thin_border

    # Results
    ws.cell(row=16, column=1, value="PERSONAL NECESARIO POR TURNO").font = section_font

    # Covers per service (if 2 services, split)
    ws.cell(row=17, column=1, value="Covers por servicio:").font = bold_font
    ws.cell(row=17, column=1).border = thin_border
    ws.cell(row=17, column=2, value='=IF(B7=3,CEILING(B5/2,1),B5)').font = formula_font
    ws.cell(row=17, column=2).alignment = center_align
    ws.cell(row=17, column=2).border = thin_border

    calc_rows = [
        ("Cocineros necesarios / turno:", 18, '=CEILING(B17/B12,1)'),
        ("Camareros necesarios / turno:", 19, '=CEILING(B17/B13,1)'),
        ("Barmans necesarios / turno:", 20, '=CEILING(B17/B14,1)'),
        ("TOTAL personal / turno:", 21, '=B18+B19+B20'),
        ("Turnos al día:", 22, '=IF(B7=3,2,1)'),
        ("Personal total necesario (con rotación):", 23, '=CEILING(B21*B22*1.15,1)'),
    ]
    for label, r, formula in calc_rows:
        ws.cell(row=r, column=1, value=label).font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=formula).font = formula_font
        ws.cell(row=r, column=2).alignment = center_align
        ws.cell(row=r, column=2).border = thin_border

    # Cost estimate
    ws.cell(row=25, column=1, value="COSTE LABORAL ESTIMADO").font = section_font
    cost_rows = [
        ("Coste mensual bruto plantilla (€):", 26, '=B23*B9'),
        ("Seguridad Social empresa (30%):", 27, '=B26*0.30'),
        ("COSTE LABORAL TOTAL MENSUAL (€):", 28, '=B26+B27'),
        ("Coste laboral anual estimado (€):", 29, '=B28*14'),
    ]
    for label, r, formula in cost_rows:
        ws.cell(row=r, column=1, value=label).font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=formula).font = formula_font
        ws.cell(row=r, column=2).number_format = '€#,##0.00'
        ws.cell(row=r, column=2).border = thin_border

    # Comparison
    ws.cell(row=31, column=1, value="COMPARATIVA CON PLANTILLA ACTUAL").font = section_font
    ws.cell(row=32, column=1, value="Personal actual:").font = bold_font
    ws.cell(row=32, column=1).border = thin_border
    ws.cell(row=32, column=2).fill = input_fill
    ws.cell(row=32, column=2).border = thin_border
    ws.cell(row=32, column=2).alignment = center_align

    ws.cell(row=33, column=1, value="Diferencia:").font = bold_font
    ws.cell(row=33, column=1).border = thin_border
    ws.cell(row=33, column=2, value='=IF(B32<>"",B32-B23,"")').font = formula_font
    ws.cell(row=33, column=2).alignment = center_align
    ws.cell(row=33, column=2).border = thin_border

    ws.cell(row=34, column=1, value="Diagnóstico:").font = bold_font
    ws.cell(row=34, column=1).border = thin_border
    ws.cell(row=34, column=2, value='=IF(B33="","",IF(B33>1,"⚠ SOBREDIMENSIONADA (+"&B33&")",IF(B33<-1,"🔴 INFRADIMENSIONADA ("&B33&")","✓ DIMENSIONADA CORRECTAMENTE")))').font = formula_font
    ws.cell(row=34, column=2).border = thin_border

    add_brand_footer(ws, 37, "D")

    # --- Sheet 3: Ratios por Tipo ---
    ws2 = wb.create_sheet("Ratios por Tipo")
    ws2.sheet_properties.tabColor = "BF360C"
    set_col_widths(ws2, [("A", 28), ("B", 20), ("C", 20), ("D", 20), ("E", 22)])

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Ratios de Personal por Tipo de Negocio"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:E2")
    ws2["A2"].value = BRAND_LINE
    ws2["A2"].font = subtitle_font

    ref_headers = ["Tipo de Negocio", "Covers/Cocinero", "Covers/Camarero", "Covers/Barman", "Ratio Coste Laboral"]
    style_header_row(ws2, 4, ref_headers)

    ref_data = [
        ("Restaurante Casual", "18-22", "10-14", "20-30", "28-33%"),
        ("Fine Dining / Alta Cocina", "6-10", "4-8", "12-18", "35-40%"),
        ("Fast Casual / Comida Rápida", "25-35", "18-25", "30-40", "25-28%"),
        ("Cafetería / Brunch", "15-20", "10-15", "20-25", "28-32%"),
        ("Bar / Cocktails", "—", "15-20", "10-15", "25-30%"),
        ("Pizzería", "20-25", "12-16", "25-30", "26-30%"),
        ("Dark Kitchen / Delivery", "25-35", "—", "—", "22-28%"),
        ("Hotel (restaurante)", "12-18", "8-12", "15-20", "35-42%"),
        ("Catering / Eventos", "15-20", "8-12", "20-25", "30-35%"),
        ("Heladería / Obrador", "20-30", "15-20", "—", "25-30%"),
    ]
    for idx, row_data in enumerate(ref_data):
        r = 5 + idx
        for ci, val in enumerate(row_data):
            cell = ws2.cell(row=r, column=ci + 1, value=val)
            cell.font = data_font
            cell.alignment = center_align if ci > 0 else left_align
            cell.border = thin_border

    # Notes
    ws2.cell(row=16, column=1, value="NOTAS:").font = section_font
    notes = [
        "▸ Los ratios son orientativos — ajustar según carta, servicio y nivel de exigencia.",
        "▸ En temporada alta, reducir el ratio (más personal por cover).",
        "▸ Considerar 15% adicional para cubrir vacaciones, bajas y descansos.",
        "▸ El ratio coste laboral incluye SS empresa (~30% sobre bruto).",
    ]
    for idx, note in enumerate(notes):
        ws2.cell(row=17 + idx, column=1, value=note).font = Font(name="Calibri", size=10, color="555555")

    add_brand_footer(ws2, 23, "E")

    fname = os.path.join(OUTPUT_DIR, "BONUS-02-calculadora-plantilla-optima.xlsx")
    wb.save(fname)
    print(f"  ✅ {fname}")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("━" * 60)
    print("  Kit Gestión de Personal y Turnos — AI Chef Pro")
    print("━" * 60)
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_bonus_01()
    gen_bonus_02()
    print("━" * 60)
    print(f"  ✅ 9 archivos generados en: {OUTPUT_DIR}")
    print("━" * 60)
