#!/usr/bin/env python3
"""
Kit de Tareas Recurrentes — Pastelería / Obrador · v2.0
Generador de los ficheros 08 (reescrito) y 10, 11, 12, 13 (nuevos).

AI Chef Pro — aichef.pro/kit-tareas-pasteleria

Salida: astro-site/public/dl/kit-tareas-pasteleria/
  08-apertura-cierre-negocio.xlsx      (reescrito para pastelería con tienda + obrador)
  10-plan-produccion-semanal.xlsx      (plan semanal + control de mermas)
  11-control-encargos.xlsx             (ficha de encargo + registro + agenda)
  12-control-alergenos-vitrina.xlsx    (matriz borrador + carta + cartel + etiquetas)
  13-registro-temperaturas-recepcion.xlsx (temperaturas + recepción + etiquetas)

NO toca 01-07, 09 ni los BONUS.

Uso: python3 scripts/generate-tareas-pasteleria-extras.py
"""

import contextlib
import datetime
import logging
import math
import os
import re
import subprocess
import sys

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ══════════════════════════════════════════════════════════════════════
# Rutas
# ══════════════════════════════════════════════════════════════════════
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(ROOT, "astro-site", "public", "dl", "kit-tareas-pasteleria")
INJECT = os.path.join(ROOT, "scripts", "productos-digitales", "inject_cache.py")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════
# Marca (idéntica al generador v1.1: scripts/generate-tareas-pasteleria.py)
# ══════════════════════════════════════════════════════════════════════
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
INPUT_GREEN = "E8F5E9"

OBRADOR_COLOR = "FFF3E0"
HORNO_COLOR = "FFEBEE"
VITRINA_COLOR = "E3F2FD"
ALMACEN_COLOR = "E8F5E9"
LIMPIEZA_COLOR = "FCE4EC"
ADMIN_COLOR = "FFF8E1"
EVENTO_COLOR = "E0F2F1"
DESPACHO_COLOR = "F3E5F5"

ZONE_COLORS = {
    "Obrador": OBRADOR_COLOR,
    "Horno": HORNO_COLOR,
    "Vitrina": VITRINA_COLOR,
    "Almacén": ALMACEN_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Evento": EVENTO_COLOR,
    "Despacho": DESPACHO_COLOR,
    "Tienda": DESPACHO_COLOR,
    "Caja": ADMIN_COLOR,
    "Acceso": LIGHT_GRAY,
    "Exterior": EVENTO_COLOR,
    "Baños": LIMPIEZA_COLOR,
    "Equipo": ADMIN_COLOR,
    "General": LIGHT_GRAY,
}

title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
note_font = Font(name="Calibri", size=10, color="555555")
alert_font = Font(name="Calibri", size=11, bold=True, color="B71C1C")
gray_mark_font = Font(name="Calibri", size=12, color="9E9E9E")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
input_fill = PatternFill(start_color=INPUT_GREEN, end_color=INPUT_GREEN, fill_type="solid")
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
alert_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
amber_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
green_ok_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
neutral_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
# relleno de la fila al marcar «✓» en la columna «✓ Completada» (mismo verde en los 22 checklists)
check_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)
cut_border = Border(
    left=Side(style="medium", color="9E9E9E"),
    right=Side(style="medium", color="9E9E9E"),
    top=Side(style="medium", color="9E9E9E"),
    bottom=Side(style="medium", color="9E9E9E"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")
rot_align = Alignment(horizontal="center", vertical="bottom", textRotation=90, wrap_text=False)

EUR = '#,##0.00 €'
PCT = '0.0%'
PCT0 = '0%'  # reservado: los porcentajes del kit van a un decimal (0,0 %)

VERSION_LINE = "Versión 2.0 · agosto 2026 · aichef.pro/kit-tareas-pasteleria · info@aichef.pro"
BIO_LINE = ("Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, "
            "en cocina desde los 17 años · johnguerrero.es")
SOPORTE_LINE = "Dudas o sugerencias: info@aichef.pro — respondemos en 24-48 h laborables."
MARCA_LINE = "Marca con ✓ en la columna «✓ Completada»: es la que cuenta el total de tareas completadas."

V2_DATE = datetime.datetime(2026, 8, 21)

ALERGENOS = [
    ("Gluten", "Cereales con gluten: trigo, centeno, cebada, avena, espelta, kamut y sus derivados"),
    ("Crustáceos", "Crustáceos y productos a base de crustáceos"),
    ("Huevos", "Huevos y productos a base de huevo"),
    ("Pescado", "Pescado y productos a base de pescado"),
    ("Cacahuetes", "Cacahuetes y productos a base de cacahuete"),
    ("Soja", "Soja y productos a base de soja, incluida la lecitina de soja"),
    ("Lácteos", "Leche y sus derivados, incluida la lactosa"),
    ("Frutos de cáscara", "Almendra, avellana, nuez, anacardo, pistacho, pacana, nuez de Brasil y macadamia"),
    ("Apio", "Apio y productos derivados"),
    ("Mostaza", "Mostaza y productos derivados"),
    ("Sésamo", "Granos de sésamo y productos a base de sésamo"),
    ("Sulfitos", "Dióxido de azufre y sulfitos en concentración superior a 10 mg/kg o 10 mg/l"),
    ("Altramuces", "Altramuces y productos a base de altramuces"),
    ("Moluscos", "Moluscos y productos a base de moluscos"),
]

NOMBRES = {
    "08": "Apertura y Cierre del Negocio",
    "10": "Plan de Producción Semanal y Control de Mermas",
    "11": "Control de Encargos",
    "12": "Control de Alérgenos de Vitrina",
    "13": "Registro de Temperaturas y Recepción de Mercancía",
}


# ══════════════════════════════════════════════════════════════════════
# Helpers de marca
# ══════════════════════════════════════════════════════════════════════
def set_metadata(wb, num):
    p = wb.properties
    p.creator = "AI Chef Pro"
    p.lastModifiedBy = "AI Chef Pro"
    p.title = f"{num} — {NOMBRES[num]} · Kit de Tareas Pastelería / Obrador"
    p.subject = "Kit de Tareas Recurrentes · Pastelería / Obrador · v2.0"
    p.keywords = "pastelería, obrador, checklist, tareas, AI Chef Pro"
    p.description = "aichef.pro/kit-tareas-pasteleria"
    p.category = "AI Chef Pro · Productos digitales"
    p.created = V2_DATE
    p.modified = V2_DATE


def print_setup(ws, header_rows=None, landscape=True, one_page=False,
                gridlines=True, print_area=None):
    """A4 + ajustar al ancho + repetir cabecera + inmovilizar + pie de marca."""
    autofit_merged(ws)
    autofit_rows(ws)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1 if one_page else 0
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = "AI Chef Pro · aichef.pro · Página &P de &N"
    ws.oddFooter.center.size = 8
    ws.print_options.horizontalCentered = True
    if not gridlines:
        ws.sheet_view.showGridLines = False
    if print_area:
        ws.print_area = print_area
    if header_rows:
        first, last = header_rows
        ws.print_title_rows = f"{first}:{last}"
        ws.freeze_panes = ws.cell(row=last + 1, column=1).coordinate


def instructions_sheet(wb, title, blocks, extra_lines=None):
    """Hoja Instrucciones en columna B, idéntica a la familia «▸» del kit."""
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 95

    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    ws["B2"].alignment = left_align

    row = 4
    for line in blocks:
        cell = ws.cell(row=row, column=2, value=line if line else None)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line.startswith("!"):
            cell.value = line[1:].strip()
            cell.font = alert_font
            cell.fill = alert_fill
            cell.border = thin_border
        elif line:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")
        cell.alignment = left_align
        row += 1

    row += 1
    for line in (extra_lines or []):
        ws.cell(row=row, column=2, value=line).font = Font(name="Calibri", size=11, bold=True, color="333333")
        ws.cell(row=row, column=2).alignment = left_align
        row += 1

    row += 1
    ws.cell(row=row, column=2, value=SOPORTE_LINE).font = note_font
    row += 1
    ws.cell(row=row, column=2, value=BIO_LINE).font = note_font
    row += 1
    ws.cell(row=row, column=2, value=VERSION_LINE).font = small_font

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.oddFooter.center.text = "AI Chef Pro · aichef.pro · Página &P de &N"
    ws.oddFooter.center.size = 8
    return ws


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ── Alto de fila para texto envuelto ──────────────────────────────────
# Excel NO autoajusta el alto de una celda COMBINADA con wrap_text: se queda en
# los 15 pt por defecto (una línea) y el resto del texto no se ve ni en pantalla
# ni en papel. Toda celda combinada con wrap_text lleva alto explícito, y las no
# combinadas con texto largo también, porque el alto fijo que traían las cortaba.
LINE_PT = 15.0          # alto de una línea de Calibri 11
PAD_PT = 4.0            # margen interior de la celda


def est_lines(text, width_chars, size=11):
    """Líneas que ocupa un texto envuelto en una celda de `width_chars` de ancho.

    El ancho de columna de Excel se mide en caracteres de la fuente por defecto
    (Calibri 11), así que un texto en minúscula suele entrar más holgado: la
    estimación es deliberadamente conservadora (mejor una línea de más que un
    texto cortado).
    """
    if not text:
        return 1
    cpl = max(6, int(width_chars * 11.0 / size))
    return max(1, sum(max(1, math.ceil(len(seg) / cpl)) for seg in str(text).split("\n")))


def wrap_height(text, width_chars, size=11, min_height=LINE_PT):
    """Alto en puntos que necesita `text` en una celda de ese ancho.

    Una línea de Calibri 11 mide 15 pt (el alto por defecto de Excel), así que un
    texto de una sola línea no infla nada; a partir de dos se añade un colchón.
    """
    line = max(12.0, size * 1.32)
    n = est_lines(text, width_chars, size)
    return max(min_height, n * line + (PAD_PT if n > 1 else 0.0))


def fit_row(ws, row, text, width_chars, size=11, rows=1, min_height=LINE_PT):
    """Fija el alto de la fila (o lo reparte entre las `rows` de la combinación)."""
    total = wrap_height(text, width_chars, size, min_height)
    per = max(min_height, round(total / rows, 1))
    for r in range(row, row + rows):
        ws.row_dimensions[r].height = per
    return per


def merged_width(ws, first_col, last_col):
    """Ancho combinado (en caracteres) de un rango de columnas."""
    total = 0.0
    for i in range(first_col, last_col + 1):
        dim = ws.column_dimensions[get_column_letter(i)]
        total += dim.width if dim.width else 8.43
    return total


def autofit_merged(ws):
    """Da alto suficiente a TODA celda combinada con wrap_text de la hoja.

    Excel autoajusta las celdas sueltas, pero nunca las combinadas: sin esto,
    los textos legales y las advertencias se quedan en una línea y el resto no
    se ve ni en pantalla ni al imprimir. Solo sube el alto, nunca lo baja.
    """
    for rng in list(ws.merged_cells.ranges):
        cell = ws.cell(row=rng.min_row, column=rng.min_col)
        if not isinstance(cell.value, str) or not cell.value.strip():
            continue
        if not (cell.alignment and cell.alignment.wrap_text):
            continue
        size = cell.font.size or 11
        need = wrap_height(cell.value, merged_width(ws, rng.min_col, rng.max_col), size)
        filas = range(rng.min_row, rng.max_row + 1)
        actual = sum(ws.row_dimensions[r].height or LINE_PT for r in filas)
        if need > actual + 0.5:
            por_fila = round(need / len(filas), 1)
            for r in filas:
                ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or LINE_PT, por_fila)


def autofit_rows(ws):
    """Segunda pasada: en las filas con alto EXPLÍCITO, que quepan también las
    celdas sueltas con wrap_text.

    Sin alto explícito Excel autoajusta, pero en cuanto una celda combinada de la
    misma fila obliga a fijarlo, el rótulo de al lado se queda cortado. Se ignoran
    las fórmulas: en pantalla muestran su resultado, no su texto.
    """
    combinadas = set()
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                combinadas.add((r, c))
    for fila in ws.iter_rows():
        r = fila[0].row
        actual = ws.row_dimensions[r].height
        if not actual:
            continue
        need = actual
        for c in fila:
            if (c.row, c.column) in combinadas:
                continue
            if not isinstance(c.value, str) or not c.value.strip() or c.value.startswith("="):
                continue
            if not (c.alignment and c.alignment.wrap_text):
                continue
            ancho = ws.column_dimensions[c.column_letter].width or 8.43
            need = max(need, wrap_height(c.value, ancho, c.font.size or 11))
        if need > actual + 0.5:
            ws.row_dimensions[r].height = round(need, 1)


def table_header(ws, row, headers, rotate=None, height=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = rot_align if (rotate and i in rotate) else center_align
        c.border = thin_border
    if height:
        ws.row_dimensions[row].height = height


def sheet_title(ws, row, last_col, text, subtitle=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = title_font
    c.alignment = left_align
    ancho = merged_width(ws, 1, last_col)
    fit_row(ws, row, text, ancho, size=16, min_height=22)
    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=last_col)
        s = ws.cell(row=row + 1, column=1, value=subtitle)
        s.font = subtitle_font
        s.alignment = left_align
        fit_row(ws, row + 1, subtitle, ancho, size=11)


def paint(ws, row, col, value=None, font=None, fill=None, align=None,
          border=True, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or left_align
    if border:
        c.border = thin_border
    if fmt:
        c.number_format = fmt
    return c


def block_row(ws, row, last_col, text, fill=None, font=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font or section_font
    c.alignment = left_align
    f = fill or gray_fill
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).fill = f
        ws.cell(row=row, column=col).border = thin_border
    return c


def fill_range(ws, first_row, first_col, last_row, last_col, fill, border=None):
    """Extiende relleno (y borde) a TODAS las celdas de una combinación.

    En una celda combinada solo se estiliza la de arriba a la izquierda: el resto
    del área se queda sin fondo y el verde de «editable» no se lee como tal.
    """
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            if border:
                cell.border = border


def footer_line(ws, row, col=1):
    ws.cell(row=row, column=col,
            value="— Kit de Tareas Recurrentes · Pastelería / Obrador · AI Chef Pro · aichef.pro").font = small_font


# ══════════════════════════════════════════════════════════════════════
# 08 — Apertura y Cierre del Negocio (tienda / despacho + obrador)
# ══════════════════════════════════════════════════════════════════════
# Horario de tienda del kit (fuente única, la misma en el 01 y en el 08):
# 06:45 encender vitrinas · 07:45 montar la vitrina (60 min después, mínimo 45)
# 08:00 apertura al público · 19:50 se cierra la puerta al público.
MINUTOS_VITRINA = 45

APERTURA_NEGOCIO = [
    ("Desactivar la alarma y encender las vitrinas refrigeradas: anotar la temperatura (2-6 °C) — mínimo 45 minutos antes de montar el género (→ 13 Registro de Temperaturas · hoja Registro de Temperaturas)", "Vitrina", "Encargado", "06:45"),
    ("Encender las luces generales, el escaparate y el rótulo exterior", "General", "Encargado", "06:45"),
    ("Encender la climatización o la calefacción: la tienda entre 20 y 22 °C (por encima sufren el chocolate y el hojaldre)", "General", "Encargado", "06:50"),
    ("Encender el TPV y el datáfono; comprobar el rollo de papel y la conexión de la pasarela de pago", "Caja", "Encargado", "06:55"),
    ("Contar el fondo de caja y dejarlo anotado (→ 09 Apertura y Cierre de Caja · hoja Apertura de Caja)", "Caja", "Encargado", "07:00"),
    ("Revisar los encargos del día y separar los que ya están listos (→ 11 Control de Encargos · hoja Encargos de Hoy)", "Tienda", "Dependiente", "07:05"),
    ("Recoger la producción del obrador y contrastarla con el plan del día (→ 10 Plan de Producción Semanal · hoja Plan Semanal)", "Obrador", "Dependiente", "07:15"),
    ("Reponer el packaging del mostrador: cajas, bolsas, papel antigrasa, cintas, blondas y velas", "Tienda", "Dependiente", "07:25"),
    ("Limpiar los cristales de la vitrina, el mostrador y la puerta de entrada antes de montar el género", "Tienda", "Ayudante", "07:30"),
    ("Comprobar que las vitrinas han bajado a 2-6 °C antes de montar; si no han llegado, retrasar el montaje", "Vitrina", "Dependiente", "07:40"),
    ("Montar la vitrina en orden: primero bollería, después pastelería refrigerada y al final salado", "Vitrina", "Dependiente", "07:45"),
    ("Revisar el producto del día anterior: retirar lo que no esté en punto y anotarlo como merma (→ 10 Plan de Producción Semanal · hoja Producido vs Vendido)", "Vitrina", "Dependiente", "07:45"),
    ("Colocar la etiqueta de precio y de alérgenos de cada referencia (→ 12 Control de Alérgenos de Vitrina · hoja Etiquetas Vitrina)", "Vitrina", "Dependiente", "07:50"),
    ("Comprobar que el cartel de información de alérgenos está a la vista del cliente (→ 12 Control de Alérgenos de Vitrina · hoja Cartel Vitrina)", "Tienda", "Dependiente", "07:50"),
    ("Sacar la pizarra de novedades y actualizar la oferta del día", "Exterior", "Ayudante", "07:50"),
    ("Preparar la máquina de café y la zona de degustación si la hay: purga, molienda y leche", "Tienda", "Ayudante", "07:55"),
    ("Repasar el baño y la zona de clientes: papel, jabón, papelera y suelo", "Baños", "Ayudante", "07:55"),
    ("Poner la música ambiente a un volumen que permita hablar en el mostrador", "General", "Ayudante", "07:55"),
    ("Briefing con el equipo: encargos del día, novedades y objetivo de venta (→ BONUS-01 Briefing de Servicio)", "Equipo", "Encargado", "07:58"),
    ("Subir la persiana y abrir la puerta al público", "Acceso", "Encargado", "08:00"),
]

CIERRE_NEGOCIO = [
    ("Última hora: agrupar y señalizar el producto en promoción para venderlo antes de cerrar (→ 10 Plan de Producción Semanal · columna Acción)", "Vitrina", "Dependiente", "19:20"),
    ("Preparar y separar los encargos de mañana con su ficha grapada (→ 11 Control de Encargos · hoja Ficha de Encargo)", "Tienda", "Dependiente", "19:40"),
    ("Cerrar la puerta al público y comprobar que no queda ningún cliente dentro", "Acceso", "Encargado", "19:50"),
    ("Retirar el género de la vitrina y anotar el sobrante de cada referencia (→ 10 Plan de Producción Semanal · hoja Producido vs Vendido)", "Vitrina", "Dependiente", "19:55"),
    ("Guardar en cámara lo recuperable, tapado y etiquetado con fecha y hora (→ 13 Registro de Temperaturas · hoja Etiquetas de Elaborado)", "Vitrina", "Dependiente", "20:00"),
    ("Decidir el destino del sobrante no recuperable y registrarlo: donación o desecho (→ 10 Plan de Producción Semanal · hoja Producido vs Vendido)", "Vitrina", "Dependiente", "20:05"),
    ("Recoger la pizarra de novedades y la señalización exterior", "Exterior", "Ayudante", "20:05"),
    ("Cierre de caja: sacar la Z del TPV, arquear el efectivo y anotar el descuadre (→ 09 Apertura y Cierre de Caja · hoja Cierre de Caja)", "Caja", "Encargado", "20:10"),
    ("Separar el fondo de caja de mañana y guardar la recaudación en la caja fuerte", "Caja", "Encargado", "20:15"),
    ("Vaciar y limpiar la vitrina por dentro: bandejas, blondas, expositores y cristales", "Vitrina", "Ayudante", "20:15"),
    ("Dejar las vitrinas en modo noche o apagadas según indique el fabricante y anotar la temperatura de cierre (→ 13 Registro de Temperaturas)", "Vitrina", "Ayudante", "20:25"),
    ("Limpiar el mostrador, la máquina de café y la zona de degustación", "Tienda", "Ayudante", "20:30"),
    ("Barrer y fregar la tienda y la zona de clientes", "Tienda", "Ayudante", "20:35"),
    ("Apagar la música, la climatización, el escaparate y el rótulo", "General", "Ayudante", "20:40"),
    ("Apagar el TPV y el datáfono, o dejarlos en reposo según indique el proveedor", "Caja", "Encargado", "20:45"),
    ("Comprobaciones finales: hornos apagados, llave de gas cerrada, cámaras cerradas y a temperatura, grifos cerrados", "General", "Encargado", "20:50"),
    ("Sacar la basura y separar el reciclaje: cartón, vidrio, orgánico y aceite usado", "General", "Ayudante", "20:55"),
    ("Bajar la persiana, apagar las luces y activar la alarma", "Acceso", "Encargado", "21:00"),
    ("Cerrar con llave y comprobar el cierre desde la calle", "Acceso", "Encargado", "21:05"),
]

HOLGURA_08 = 5


def negocio_sheet(wb, name, tab_color, title, subtitle, tasks):
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = tab_color
    widths(ws, {"A": 5, "B": 62, "C": 13, "D": 14, "E": 12, "F": 14, "G": 14, "H": 24})

    sheet_title(ws, 1, 8, title, subtitle)

    headers = ["Nº", "Tarea", "Zona", "Responsable", "Hora Límite", "✓ Completada", "Firma", "Notas"]
    table_header(ws, 4, headers)

    dv = DataValidation(type="list", formula1='"✓,—,N/A"', allow_blank=True,
                        promptTitle="Estado de la tarea",
                        prompt="✓ hecha · — no aplica hoy · N/A no aplica en este local")
    ws.add_data_validation(dv)

    row = 5
    first_data = row
    for i, (tarea, zona, resp, hora) in enumerate(tasks, 1):
        paint(ws, row, 1, i, font=Font(name="Calibri", size=10, color="666666"), align=center_align)
        paint(ws, row, 2, tarea)
        zf = PatternFill(start_color=ZONE_COLORS.get(zona, LIGHT_GRAY),
                         end_color=ZONE_COLORS.get(zona, LIGHT_GRAY), fill_type="solid")
        paint(ws, row, 3, zona, font=Font(name="Calibri", size=10, color="666666"),
              fill=zf, align=center_align)
        paint(ws, row, 4, resp, fill=input_fill, align=center_align)
        paint(ws, row, 5, hora, fill=input_fill, align=center_align)
        paint(ws, row, 6, None, fill=input_fill, align=center_align)
        paint(ws, row, 7, None, fill=input_fill)
        paint(ws, row, 8, None, fill=input_fill)
        # el alto sale del texto de la tarea: con 30 pt fijos las tareas largas se cortaban
        fit_row(ws, row, tarea, ws.column_dimensions["B"].width, size=11, min_height=30)
        row += 1

    for _ in range(HOLGURA_08):
        for col in range(1, 9):
            paint(ws, row, col, None, fill=input_fill,
                  align=center_align if col != 2 else left_align)
        ws.row_dimensions[row].height = 22
        row += 1
    last_data = row - 1
    dv.add(f"F{first_data}:F{last_data}")

    # la fila entera se pinta de verde al marcar «✓» en la columna «✓ Completada»
    ws.conditional_formatting.add(
        f"A{first_data}:H{last_data}",
        FormulaRule(formula=[f'$F{first_data}="✓"'], fill=check_fill))

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="Tareas completadas:").font = bold_font
    ws.cell(row=row, column=1).alignment = left_align
    c = ws.cell(row=row, column=4, value=f'=COUNTIF(F{first_data}:F{last_data},"✓")')
    c.font = bold_font
    c.alignment = center_align
    ws.cell(row=row, column=5, value="de").font = data_font
    ws.cell(row=row, column=5).alignment = center_align
    c = ws.cell(row=row, column=6, value=f'=COUNTIF(B{first_data}:B{last_data},"?*")')
    c.font = bold_font
    c.alignment = center_align
    contador_row = row

    row += 2
    ws.cell(row=row, column=1, value="Verificado por:").font = bold_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    paint(ws, row, 2, None, fill=input_fill)
    ws.cell(row=row, column=5, value="Firma:").font = bold_font
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
    paint(ws, row, 6, None, fill=input_fill)

    row += 2
    footer_line(ws, row)

    print_setup(ws, header_rows=(4, 4), landscape=True)
    return contador_row


def generate_08():
    wb = Workbook()
    instructions_sheet(
        wb,
        "08 — Apertura y Cierre del Negocio · Pastelería con Tienda y Obrador",
        [
            "Qué cubre esta plantilla:",
            "▸ El LOCAL: tienda, despacho, vitrinas, caja y accesos. La producción del obrador va en el fichero 01.",
            "▸ Las tareas están en el orden real en el que se hacen, con su hora límite orientativa.",
            "▸ La primera tarea del día son las vitrinas: enciéndelas mínimo 45 minutos antes de montar el "
            "género, que es lo que tardan en bajar a 2-6 °C. El cronograma las enciende a las 06:45, monta "
            "a las 07:45 y abre al público a las 08:00.",
            "▸ El cierre empieza vendiendo, no recogiendo: primero se agrupa y se señaliza el producto en "
            "promoción, a las 19:50 se cierra la puerta al público y solo entonces se vacía la vitrina.",
            "▸ Las cuatro últimas del cierre no se mueven de sitio: comprobaciones finales, basura, alarma y llave.",
            "",
            "Cómo usar esta plantilla:",
            "▸ Imprime la hoja de apertura y la de cierre cada día, o rellénalas en tablet.",
            "▸ El responsable del turno marca cada tarea al completarla y firma al terminar.",
            "▸ Las tareas que remiten a otro fichero del kit lo citan por su número: sigue la referencia.",
            "",
            "Personalización:",
            "▸ Las celdas verdes son editables: ajusta responsables, horas y notas a tu local.",
            "▸ Al final de cada hoja hay cinco filas libres para tus propias tareas; el contador ya las incluye.",
            "▸ Borra las tareas que no apliquen a tu local (zona de degustación, máquina de café, reparto).",
        ],
        extra_lines=[MARCA_LINE],
    )
    negocio_sheet(wb, "Apertura del Negocio", "FFD700",
                  "Apertura del Negocio — Pastelería con Tienda",
                  "Fecha: ____/____/________     Responsable de turno: _________________________     "
                  "Temperatura de vitrina al abrir: ______ °C",
                  APERTURA_NEGOCIO)
    negocio_sheet(wb, "Cierre del Negocio", "8B4513",
                  "Cierre del Negocio — Pastelería con Tienda",
                  "Fecha: ____/____/________     Responsable de turno: _________________________     "
                  "Temperatura de vitrina al cerrar: ______ °C",
                  CIERRE_NEGOCIO)
    set_metadata(wb, "08")
    path = os.path.join(OUTPUT_DIR, "08-apertura-cierre-negocio.xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════
# 10 — Plan de Producción Semanal + Control de Mermas
# ══════════════════════════════════════════════════════════════════════
CATALOGO = [
    ("Bollería", [
        "Croissant de mantequilla", "Pain au chocolat", "Napolitana de crema",
        "Ensaimada", "Brioche", "Caracola de pasas", "Palmera de hojaldre",
        "Croissant relleno de crema",
    ]),
    ("Pastelería", [
        "Tarta de queso", "Tarta de manzana", "Milhojas", "Éclair",
        "Tarta de zanahoria", "Tarta Sacher", "Entremet individual",
        "Macarons (caja de 6)", "Tartaleta de frutas", "Lionesa o profiterol",
    ]),
    ("Pan", [
        "Barra", "Hogaza de masa madre", "Chapata", "Pan de molde", "Focaccia",
    ]),
    ("Salado", [
        "Quiche", "Empanada", "Croissant de jamón y queso",
    ]),
    ("Bizcochería", [
        "Magdalena", "Cookie de chocolate", "Brownie", "Galleta de mantequilla",
        "Bizcocho de yogur", "Financier de almendra",
    ]),
]

DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

ACCIONES = ("Vender −30 % última hora,"
            "Congelar (SOLO producto no expuesto del día y abatido antes de −18 °C),"
            "Reprocesar (SOLO producto seco de menos de 24 h; registrar lote destino),"
            "Donar,"
            "Desechar")

NOTA_APPCC = ("Criterio APPCC: el producto que ha estado expuesto en vitrina NO se congela para venderlo después. "
              "Solo se congela producto del día que no ha salido a la venta y que se abate antes de bajar de −18 °C. "
              "Si vendes un producto que ha sido congelado, tienes que informar de que está descongelado. "
              "Prioriza la donación sobre el desecho: revisa el plan de prevención de pérdidas y desperdicio "
              "alimentario que te aplique antes de dar salida al sobrante.")


def gen10_plan(wb):
    ws = wb.create_sheet("Plan Semanal")
    ws.sheet_properties.tabColor = "FFD700"
    widths(ws, {"A": 30, "B": 15, "C": 8, "D": 8, "E": 8, "F": 8, "G": 8, "H": 8,
                "I": 8, "J": 15, "K": 34})
    sheet_title(ws, 1, 11, "Plan de Producción Semanal",
                "Semana del ____/____ al ____/____     Responsable: _________________________     "
                "Anota las unidades previstas por día")
    table_header(ws, 4, ["Producto", "Partida"] + DIAS + ["Total semana", "Notas"])

    row = 5
    first_data = row
    for partida, productos in CATALOGO:
        block_row(ws, row, 11, f"  {partida}",
                  fill=PatternFill(start_color=ZONE_COLORS.get("Obrador"),
                                   end_color=ZONE_COLORS.get("Obrador"), fill_type="solid"))
        row += 1
        for prod in productos:
            paint(ws, row, 1, prod)
            paint(ws, row, 2, partida, font=Font(name="Calibri", size=10, color="666666"),
                  align=center_align)
            for i in range(3, 10):
                paint(ws, row, i, None, fill=input_fill, align=center_align, fmt="0")
            c = paint(ws, row, 10, f"=SUM(C{row}:I{row})", font=bold_font,
                      align=center_align, fmt="0")
            paint(ws, row, 11, None, fill=input_fill)
            row += 1
        for _ in range(2):
            paint(ws, row, 1, None, fill=input_fill)
            paint(ws, row, 2, None, fill=input_fill, align=center_align)
            for i in range(3, 10):
                paint(ws, row, i, None, fill=input_fill, align=center_align, fmt="0")
            paint(ws, row, 10, f"=SUM(C{row}:I{row})", font=bold_font, align=center_align, fmt="0")
            paint(ws, row, 11, None, fill=input_fill)
            row += 1
    last_data = row - 1

    paint(ws, row, 1, "TOTAL DE UNIDADES", font=bold_font, fill=gold_fill)
    paint(ws, row, 2, None, fill=gold_fill)
    for i in range(3, 11):
        col = get_column_letter(i)
        paint(ws, row, i, f"=SUM({col}{first_data}:{col}{last_data})",
              font=bold_font, fill=gold_fill, align=center_align, fmt="0")
    paint(ws, row, 11, None, fill=gold_fill)
    row += 2
    ws.cell(row=row, column=1,
            value="Las filas verdes vacías de cada partida son para tus referencias: ya están dentro de los totales.").font = note_font
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)


DIARIO_EJEMPLO = [
    ("Croissant de mantequilla", "Bollería"),
    ("Pain au chocolat", "Bollería"),
    ("Napolitana de crema", "Bollería"),
    ("Palmera de hojaldre", "Bollería"),
    ("Tarta de queso", "Pastelería"),
    ("Milhojas", "Pastelería"),
    ("Tartaleta de frutas", "Pastelería"),
    ("Barra", "Pan"),
    ("Hogaza de masa madre", "Pan"),
    ("Quiche", "Salado"),
    ("Magdalena", "Bizcochería"),
    ("Cookie de chocolate", "Bizcochería"),
]
# Una línea por producto y DÍA: con 12 referencias, 40 filas se agotaban en tres
# días y medio y el Resumen «semanal» nunca llegaba a serlo. 250 filas cubren
# entre tres y ocho semanas según el número de referencias que registres.
DIARIO_FILAS = 250


def gen10_diario(wb):
    ws = wb.create_sheet("Producido vs Vendido")
    ws.sheet_properties.tabColor = "FF8C00"
    widths(ws, {"A": 12, "B": 28, "C": 14, "D": 10, "E": 11, "F": 10, "G": 11,
                "H": 10, "I": 12, "J": 14, "K": 44, "L": 28})
    sheet_title(ws, 1, 12, "Producido vs Vendido — Control Diario de Mermas",
                "Anota cada día lo que has producido y lo que has vendido: el sobrante y su coste se calculan solos")
    table_header(ws, 4, ["Fecha", "Producto", "Partida", "Previsto", "Producido", "Vendido",
                         "Sobrante", "Merma %", "Coste ud (€)", "Coste merma (€)",
                         "Acción", "Notas"])

    dv_accion = DataValidation(type="list", formula1=f'"{ACCIONES}"', allow_blank=True)
    ws.add_data_validation(dv_accion)
    dv_partida = DataValidation(type="list",
                                formula1='"Bollería,Pastelería,Pan,Salado,Bizcochería"',
                                allow_blank=True)
    ws.add_data_validation(dv_partida)

    row = 5
    first = row
    for i in range(DIARIO_FILAS):
        prod, part = DIARIO_EJEMPLO[i] if i < len(DIARIO_EJEMPLO) else (None, None)
        paint(ws, row, 1, None, fill=input_fill, align=center_align, fmt="dd/mm/yyyy")
        paint(ws, row, 2, prod, fill=input_fill)
        paint(ws, row, 3, part, fill=input_fill, align=center_align)
        # cantidades VACÍAS, no 0: un 0 hacía que la merma calculara 0 % y el
        # semáforo pintase de verde «todo correcto» sin un solo dato introducido
        for col in (4, 5, 6):
            paint(ws, row, col, None, fill=input_fill, align=center_align, fmt="0")
        paint(ws, row, 7, f"=IFERROR(E{row}-F{row},0)", align=center_align, fmt="0")
        paint(ws, row, 8, f"=IFERROR(G{row}/E{row},0)", align=center_align, fmt=PCT)
        paint(ws, row, 9, None, fill=input_fill, align=center_align, fmt=EUR)
        paint(ws, row, 10, f"=IFERROR(G{row}*I{row},0)", align=center_align, fmt=EUR)
        paint(ws, row, 11, None, fill=input_fill)
        paint(ws, row, 12, None, fill=input_fill)
        row += 1
    last = row - 1
    dv_partida.add(f"C{first}:C{last}")
    dv_accion.add(f"K{first}:K{last}")

    # Semáforo de merma CON GUARDA: sin producción anotada la fila no se pinta
    # (antes, una fila vacía daba merma 0 % y salía en verde «todo correcto»).
    ws.conditional_formatting.add(
        f"H{first}:H{last}",
        FormulaRule(formula=[f"AND($E{first}>0,$H{first}>0.1)"], fill=red_fill))
    ws.conditional_formatting.add(
        f"H{first}:H{last}",
        FormulaRule(formula=[f"AND($E{first}>0,$H{first}>=0.05,$H{first}<=0.1)"], fill=amber_fill))
    ws.conditional_formatting.add(
        f"H{first}:H{last}",
        FormulaRule(formula=[f"AND($E{first}>0,$H{first}<0.05)"], fill=green_ok_fill))

    paint(ws, row, 1, "TOTALES", font=bold_font, fill=gold_fill)
    paint(ws, row, 2, None, fill=gold_fill)
    paint(ws, row, 3, None, fill=gold_fill)
    for col in (4, 5, 6, 7):
        letter = get_column_letter(col)
        paint(ws, row, col, f"=SUM({letter}{first}:{letter}{last})",
              font=bold_font, fill=gold_fill, align=center_align, fmt="0")
    paint(ws, row, 8, f"=IFERROR(G{row}/E{row},0)", font=bold_font, fill=gold_fill,
          align=center_align, fmt=PCT)
    paint(ws, row, 9, None, fill=gold_fill)
    paint(ws, row, 10, f"=SUM(J{first}:J{last})", font=bold_font, fill=gold_fill,
          align=center_align, fmt=EUR)
    paint(ws, row, 11, None, fill=gold_fill)
    paint(ws, row, 12, None, fill=gold_fill)
    totales_row = row

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=12)
    c = ws.cell(row=row, column=1, value=NOTA_APPCC)
    c.font = note_font
    c.alignment = left_align
    c.fill = amber_fill
    c.border = thin_border
    row += 4
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)
    return first, last, totales_row


def gen10_resumen(wb, first, last):
    ws = wb.create_sheet("Resumen por Partida")
    ws.sheet_properties.tabColor = "8B4513"
    widths(ws, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 12, "F": 16, "G": 34})
    sheet_title(ws, 1, 7, "Resumen Semanal por Partida",
                "Se calcula solo a partir de la hoja Producido vs Vendido")
    table_header(ws, 4, ["Partida", "Producido", "Vendido", "Sobrante", "Merma %",
                         "Coste merma (€)", "Lectura"])

    src = "'Producido vs Vendido'"
    partidas = ["Bollería", "Pastelería", "Pan", "Salado", "Bizcochería"]
    row = 5
    firstp = row
    # el objetivo de merma editable vive tres filas por debajo de la fila TOTAL
    obj = firstp + len(partidas) + 3
    for p in partidas:
        paint(ws, row, 1, p, font=bold_font)
        paint(ws, row, 2, f'=SUMIF({src}!$C${first}:$C${last},$A{row},{src}!$E${first}:$E${last})',
              align=center_align, fmt="0")
        paint(ws, row, 3, f'=SUMIF({src}!$C${first}:$C${last},$A{row},{src}!$F${first}:$F${last})',
              align=center_align, fmt="0")
        paint(ws, row, 4, f'=SUMIF({src}!$C${first}:$C${last},$A{row},{src}!$G${first}:$G${last})',
              align=center_align, fmt="0")
        paint(ws, row, 5, f"=IFERROR(D{row}/B{row},0)", align=center_align, fmt=PCT)
        paint(ws, row, 6, f'=SUMIF({src}!$C${first}:$C${last},$A{row},{src}!$J${first}:$J${last})',
              align=center_align, fmt=EUR)
        paint(ws, row, 7, f'=IF(E{row}<=$B${obj},"OK","Revisar")', align=center_align)
        row += 1
    lastp = row - 1

    paint(ws, row, 1, "TOTAL", font=bold_font, fill=gold_fill)
    for col in (2, 3, 4):
        letter = get_column_letter(col)
        paint(ws, row, col, f"=SUM({letter}{firstp}:{letter}{lastp})",
              font=bold_font, fill=gold_fill, align=center_align, fmt="0")
    paint(ws, row, 5, f"=IFERROR(D{row}/B{row},0)", font=bold_font, fill=gold_fill,
          align=center_align, fmt=PCT)
    paint(ws, row, 6, f"=SUM(F{firstp}:F{lastp})", font=bold_font, fill=gold_fill,
          align=center_align, fmt=EUR)
    paint(ws, row, 7, f'=IF(E{row}<=$B${obj},"OK","Revisar")', font=bold_font,
          fill=gold_fill, align=center_align)
    total_row = row

    row += 2
    block_row(ws, row, 7, "  Objetivo de Merma")
    row += 1
    objetivo_row = row
    paint(ws, row, 1, "Objetivo de merma (editable)", font=bold_font)
    # mismo formato que las cifras que compara (0,0 %): con «0%» un objetivo de
    # 4,5 % se mostraba como «5%» y parecía que no se había guardado
    paint(ws, row, 2, 0.05, fill=input_fill, align=center_align, fmt=PCT)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    paint(ws, row, 3, "Escribe aquí tu objetivo: la columna Lectura compara cada partida contra él.",
          font=note_font)
    row += 1
    paint(ws, row, 1, "Merma global de la semana", font=bold_font)
    paint(ws, row, 2, f"=E{total_row}", font=bold_font, align=center_align, fmt=PCT)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    paint(ws, row, 3, f'=IF(B{row}<=B{objetivo_row},"Dentro del objetivo","Por encima del objetivo: revisa el plan")',
          font=bold_font)
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=7)
    c = ws.cell(row=row, column=1,
                value=("Criterio: la merma de una pastelería artesana suele moverse entre el 3 % y el 8 % de la "
                       "producción. Por encima del 10 % hay que revisar el plan de producción: casi siempre es "
                       "exceso de unidades en dos o tres referencias, no un problema general."))
    c.font = note_font
    c.alignment = left_align
    c.fill = amber_fill
    c.border = thin_border
    row += 3
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)

    assert objetivo_row == obj, f"objetivo_row={objetivo_row} obj={obj}"


def generate_10():
    wb = Workbook()
    instructions_sheet(
        wb,
        "10 — Plan de Producción Semanal y Control de Mermas",
        [
            "Qué resuelve esta plantilla:",
            "▸ Cuánto produces de cada referencia, cuánto vendes y cuánto tiras — con el coste en euros.",
            "▸ La merma es el margen que se te va sin que lo veas: medirla una semana ya cambia el plan.",
            "",
            "Cómo usar esta plantilla:",
            "▸ Hoja Plan Semanal: escribe las unidades previstas por día. Es el plan que baja al obrador.",
            "▸ Hoja Producido vs Vendido: al cerrar la tienda anota producido y vendido de cada referencia.",
            "▸ El sobrante, el porcentaje de merma y su coste se calculan solos. Rellena el coste por unidad "
            "con el escandallo real (si no lo tienes, empieza por el coste de materia prima).",
            "▸ Hoja Resumen por Partida: mira dónde se concentra la merma antes de tocar el plan.",
            "",
            "Columna Acción: qué hacer con el sobrante",
            "▸ Cada opción lleva su condición dentro. Léela antes de elegir: no todo el sobrante se puede congelar.",
            "▸ Registra siempre lo que decides: es lo que te permite defender tu criterio ante una inspección.",
            "",
            "Personalización:",
            "▸ Las celdas verdes son editables. Cambia los productos por los tuyos: la partida se elige de la lista.",
            "▸ Cada partida del Plan Semanal tiene dos filas libres, ya incluidas en los totales.",
            "▸ Referencias del kit: el sobrante que anotes aquí sale del cierre del fichero 08 "
            "(→ 08 Apertura y Cierre del Negocio · hoja Cierre del Negocio).",
        ],
    )
    gen10_plan(wb)
    first, last, _ = gen10_diario(wb)
    gen10_resumen(wb, first, last)
    set_metadata(wb, "10")
    path = os.path.join(OUTPUT_DIR, "10-plan-produccion-semanal.xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════
# 11 — Ficha de Encargo + Registro de Encargos
# ══════════════════════════════════════════════════════════════════════
RGPD_TITULO = "Información Básica sobre Protección de Datos"
RGPD_TEXTO = (
    "Responsable: el establecimiento indicado en esta ficha. "
    "Finalidad: gestionar, elaborar y entregar este encargo, y emitir su factura. "
    "Legitimación: ejecución del contrato solicitado por el cliente. "
    "Categorías especiales: los datos sobre alergias e intolerancias se tratan con su consentimiento "
    "explícito (art. 9.2.a RGPD) y con la única finalidad de elaborar el encargo con seguridad. "
    "Datos de terceros: si facilita datos de otras personas (comensales), declara contar con su autorización. "
    "Conservación: mientras dure el encargo y, después, durante los plazos legales de facturación. "
    "Destinatarios: no se ceden datos a terceros salvo obligación legal. "
    "Derechos: acceso, rectificación, supresión, oposición, limitación y portabilidad escribiendo a "
    "______________________________. También puede presentar una reclamación ante la Agencia Española "
    "de Protección de Datos (www.aepd.es). "
    "Información adicional disponible en ______________________________."
)
RGPD_RESPONSABLE = "Responsable del tratamiento"
RGPD_SALUD = ("☐  Consiento expresamente el tratamiento de los datos de salud (alergias e intolerancias) "
              "recogidos en esta ficha, con la única finalidad de elaborar el encargo con seguridad "
              "(art. 9.2.a RGPD). Sin este consentimiento no se puede aceptar un encargo con alergias.")
RGPD_TERCEROS = ("Si facilita datos de otras personas (comensales), declara contar con su autorización "
                 "para comunicárnoslos con esta finalidad.")
RGPD_CONSENT = ("☐  Acepto recibir comunicaciones comerciales del establecimiento (novedades y campañas). "
                "Marcar esta casilla es voluntario, es independiente del consentimiento anterior y no "
                "condiciona el encargo.")

ESTADOS = '"Pendiente,En producción,Listo,Entregado,Cancelado"'
PAGOS = '"Efectivo,Tarjeta,Bizum,Transferencia,Pendiente de pago"'
CANALES = '"Tienda,Teléfono,WhatsApp,Web,Otro"'


def gen11_ficha(wb):
    ws = wb.create_sheet("Ficha de Encargo")
    ws.sheet_properties.tabColor = "FFD700"
    widths(ws, {"A": 26, "B": 22, "C": 26, "D": 22})
    ws.sheet_view.showGridLines = False

    dv_estado = DataValidation(type="list", formula1=ESTADOS, allow_blank=True)
    dv_pago = DataValidation(type="list", formula1=PAGOS, allow_blank=True)
    dv_canal = DataValidation(type="list", formula1=CANALES, allow_blank=True)
    for dv in (dv_estado, dv_pago, dv_canal):
        ws.add_data_validation(dv)

    sheet_title(ws, 1, 4, "Ficha de Encargo",
                "Rellena una ficha por encargo. Grápala al producto el día de la entrega.")

    row = 4

    def campo(label, col, value=None, dv=None, fmt=None, height=20):
        paint(ws, row, col, label, font=bold_font, fill=gray_fill)
        c = paint(ws, row, col + 1, value, fill=input_fill, fmt=fmt)
        if dv:
            dv.add(c)
        # dos campos por fila: el alto lo manda el rótulo más largo de los dos
        ancho = ws.column_dimensions[get_column_letter(col)].width or 8.43
        alto = max(height, wrap_height(label, ancho, 11))
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, alto)
        return c

    block_row(ws, row, 4, "  Datos del Encargo")
    row += 1
    campo("Nº de encargo", 1); campo("Fecha del pedido", 3, fmt="dd/mm/yyyy"); row += 1
    campo("Canal del pedido", 1, dv=dv_canal); campo("Pedido tomado por", 3); row += 1

    block_row(ws, row, 4, "  Cliente")
    row += 1
    campo("Cliente (nombre y apellidos)", 1); campo("Teléfono", 3); row += 1
    campo("Correo electrónico", 1); campo("Estado del encargo", 3, dv=dv_estado); row += 1

    block_row(ws, row, 4, "  Entrega")
    row += 1
    campo("Recogida en tienda", 1, "☐"); campo("Entrega a domicilio", 3, "☐"); row += 1
    campo("Dirección de entrega", 1); campo("Transporte y montaje in situ", 3, "☐"); row += 1
    campo("Fecha de entrega", 1, fmt="dd/mm/yyyy"); campo("Hora de entrega", 3); row += 1
    campo("Hora de salida del obrador", 1); campo("Entregado por", 3); row += 1

    block_row(ws, row, 4, "  Producto")
    row += 1
    campo("Producto o descripción", 1, height=30); campo("Raciones", 3); row += 1
    campo("Tamaño (cm)", 1); campo("Nº de pisos", 3); row += 1
    campo("Bizcocho o base", 1); campo("Relleno por piso", 3); row += 1
    campo("Cobertura y acabado", 1); campo("Boceto o foto adjunta", 3, "☐ Sí   ☐ No"); row += 1
    campo("Dedicatoria (EN MAYÚSCULAS Y LETRA CLARA)", 1, height=30)
    campo("Vela, topper o figura", 3); row += 1
    campo("¿Los aporta el cliente?", 1, "☐ Sí   ☐ No"); campo("Preparado por", 3); row += 1

    block_row(ws, row, 4, "  Alérgenos e Intolerancias del Comensal")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1,
                value=("Marca lo que el cliente declare. Consulta la ficha de cada elaboración antes de aceptar "
                       "el encargo (→ 12 Control de Alérgenos de Vitrina · hoja Carta de Alérgenos)."))
    c.font = note_font
    c.alignment = left_align
    row += 1
    for i in range(0, 14, 2):
        for j, col in ((0, 1), (1, 3)):
            if i + j < 14:
                nombre = ALERGENOS[i + j][0]
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                paint(ws, row, col, f"☐  {nombre}", fill=input_fill)
        ws.row_dimensions[row].height = 18
        row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1,
                value="☐  ALERGIA CONFIRMADA de un comensal (no es una preferencia): avisa al obrador por escrito.")
    c.font = alert_font
    c.fill = alert_fill
    c.border = thin_border
    c.alignment = left_align
    ws.row_dimensions[row].height = 22
    row += 1
    # Las alergias son datos de SALUD (categoría especial del art. 9 RGPD): su
    # consentimiento va aquí, separado del de comunicaciones comerciales.
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=RGPD_SALUD)
    c.font = Font(name="Calibri", size=9, color="555555")
    c.alignment = left_align
    c.border = thin_border
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=RGPD_TERCEROS)
    c.font = Font(name="Calibri", size=9, color="555555")
    c.alignment = left_align
    row += 1
    campo("Observaciones", 1, height=34)
    campo("Otras intolerancias", 3); row += 1

    block_row(ws, row, 4, "  Importes")
    row += 1
    total_row = row
    campo("Precio total (€, IVA incluido)", 1, 0, fmt=EUR)
    campo("Forma de pago", 3, dv=dv_pago); row += 1
    senal_row = row
    campo("Señal entregada (€, a cuenta del total)", 1, 0, fmt=EUR)
    campo("Fecha de la señal", 3, fmt="dd/mm/yyyy"); row += 1
    paint(ws, row, 1, "Pendiente de cobro (€)", font=bold_font, fill=gray_fill)
    paint(ws, row, 2, f"=IFERROR(B{total_row}-B{senal_row},0)", font=bold_font, fmt=EUR)
    paint(ws, row, 3, "Cobrado por", font=bold_font, fill=gray_fill)
    paint(ws, row, 4, None, fill=input_fill)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1,
                value="Política de señal: ______________________________________________ "
                      "(por ejemplo: la señal no se devuelve si el encargo se anula con menos de 48 h).")
    c.font = note_font
    c.alignment = left_align
    row += 2

    block_row(ws, row, 4, "  Conformidad")
    row += 1
    campo("Firma del cliente", 1, height=34)
    campo("Fecha", 3, fmt="dd/mm/yyyy"); row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=RGPD_TITULO)
    c.font = Font(name="Calibri", size=10, bold=True, color="333333")
    c.alignment = left_align
    row += 1
    # el responsable lo escribe cada establecimiento: celda verde editable
    paint(ws, row, 1, RGPD_RESPONSABLE, font=Font(name="Calibri", size=9, bold=True, color="333333"),
          fill=gray_fill)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    paint(ws, row, 2, None, fill=input_fill)
    ws.row_dimensions[row].height = 20
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=4)
    c = ws.cell(row=row, column=1, value=RGPD_TEXTO)
    c.font = Font(name="Calibri", size=9, color="555555")
    c.alignment = left_align
    c.border = thin_border
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=RGPD_CONSENT)
    c.font = Font(name="Calibri", size=9, color="555555")
    c.alignment = left_align
    row += 2
    footer_line(ws, row)

    print_setup(ws, header_rows=None, landscape=False, one_page=True,
                gridlines=False, print_area=f"A1:D{row}")


REGISTRO_FILAS = 40


def gen11_registro(wb):
    ws = wb.create_sheet("Registro de Encargos")
    ws.sheet_properties.tabColor = "FF8C00"
    widths(ws, {"A": 8, "B": 13, "C": 24, "D": 14, "E": 30, "F": 10, "G": 13, "H": 8,
                "I": 24, "J": 12, "K": 12, "L": 13, "M": 15, "N": 15, "O": 26})
    sheet_title(ws, 1, 15, "Registro de Encargos",
                "Mes: ________________     Año: ________     "
                "Una línea por encargo. Los importes pendientes se calculan solos.")
    table_header(ws, 4, ["Nº", "Fecha pedido", "Cliente", "Teléfono", "Producto", "Raciones",
                         "Fecha entrega", "Hora", "Alérgenos declarados", "Total €", "Señal €",
                         "Pendiente €", "Estado", "Responsable", "Notas"])

    dv_estado = DataValidation(type="list", formula1=ESTADOS, allow_blank=True)
    ws.add_data_validation(dv_estado)

    row = 5
    first = row
    for i in range(1, REGISTRO_FILAS + 1):
        paint(ws, row, 1, i, font=Font(name="Calibri", size=10, color="666666"), align=center_align)
        paint(ws, row, 2, None, fill=input_fill, align=center_align, fmt="dd/mm/yyyy")
        paint(ws, row, 3, None, fill=input_fill)
        paint(ws, row, 4, None, fill=input_fill, align=center_align)
        paint(ws, row, 5, None, fill=input_fill)
        paint(ws, row, 6, None, fill=input_fill, align=center_align, fmt="0")
        paint(ws, row, 7, None, fill=input_fill, align=center_align, fmt="dd/mm/yyyy")
        paint(ws, row, 8, None, fill=input_fill, align=center_align)
        paint(ws, row, 9, None, fill=input_fill)
        paint(ws, row, 10, 0, fill=input_fill, align=center_align, fmt=EUR)
        paint(ws, row, 11, 0, fill=input_fill, align=center_align, fmt=EUR)
        paint(ws, row, 12, f"=IFERROR(J{row}-K{row},0)", align=center_align, fmt=EUR)
        paint(ws, row, 13, None, fill=input_fill, align=center_align)
        paint(ws, row, 14, None, fill=input_fill, align=center_align)
        paint(ws, row, 15, None, fill=input_fill)
        row += 1
    last = row - 1
    dv_estado.add(f"M{first}:M{last}")

    for estado, fill in (("Listo", green_ok_fill), ("Entregado", neutral_fill),
                         ("Cancelado", red_fill), ("En producción", amber_fill)):
        ws.conditional_formatting.add(
            f"M{first}:M{last}",
            CellIsRule(operator="equal", formula=[f'"{estado}"'], fill=fill))

    paint(ws, row, 1, "TOTALES", font=bold_font, fill=gold_fill)
    for col in range(2, 10):
        paint(ws, row, col, None, fill=gold_fill)
    for col in (10, 11, 12):
        letter = get_column_letter(col)
        paint(ws, row, col, f"=SUM({letter}{first}:{letter}{last})", font=bold_font,
              fill=gold_fill, align=center_align, fmt=EUR)
    for col in (13, 14, 15):
        paint(ws, row, col, None, fill=gold_fill)
    row += 1
    paint(ws, row, 1, "Encargos", font=bold_font)
    paint(ws, row, 2, f'=COUNTIF(C{first}:C{last},"?*")', font=bold_font, align=center_align, fmt="0")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=15)
    paint(ws, row, 3, "Encargos con cliente anotado en el mes.", font=note_font)
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)


def gen11_agenda(wb):
    ws = wb.create_sheet("Agenda de Entregas")
    ws.sheet_properties.tabColor = "8B4513"
    widths(ws, {"A": 10, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 24})
    sheet_title(ws, 1, 8, "Agenda de Entregas de la Semana",
                "Semana del ____/____ al ____/____     Anota «cliente · producto» en su franja")
    table_header(ws, 4, ["Hora"] + DIAS)
    row = 5
    for h in range(8, 21):
        paint(ws, row, 1, f"{h:02d}:00", font=bold_font, fill=gray_fill, align=center_align)
        for col in range(2, 9):
            paint(ws, row, col, None, fill=input_fill)
        ws.row_dimensions[row].height = 26
        row += 1
    row += 1
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)


def gen11_hoy(wb):
    ws = wb.create_sheet("Encargos de Hoy")
    ws.sheet_properties.tabColor = "FFD700"
    widths(ws, {"A": 8, "B": 10, "C": 26, "D": 34, "E": 14, "F": 12, "G": 26})
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 7, "Encargos de Hoy",
                "Fecha: ____/____/________     Imprime esta hoja y cuélgala en el obrador y en el mostrador")
    table_header(ws, 4, ["Nº", "Hora", "Cliente", "Producto", "Estado", "Cobrado", "Observaciones"])
    dv_estado = DataValidation(type="list", formula1=ESTADOS, allow_blank=True)
    ws.add_data_validation(dv_estado)
    row = 5
    for i in range(1, 17):
        paint(ws, row, 1, i, font=Font(name="Calibri", size=10, color="666666"), align=center_align)
        paint(ws, row, 2, None, fill=input_fill, align=center_align)
        paint(ws, row, 3, None, fill=input_fill)
        paint(ws, row, 4, None, fill=input_fill)
        paint(ws, row, 5, None, fill=input_fill, align=center_align)
        paint(ws, row, 6, None, fill=input_fill, align=center_align)
        paint(ws, row, 7, None, fill=input_fill)
        ws.row_dimensions[row].height = 24
        row += 1
    dv_estado.add(f"E5:E{row - 1}")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1,
            value="Cada encargo sale del obrador con su ficha grapada (→ hoja Ficha de Encargo).").font = note_font
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=False, one_page=True,
                gridlines=False, print_area=f"A1:G{row}")


def generate_11():
    wb = Workbook()
    instructions_sheet(
        wb,
        "11 — Control de Encargos · Pastelería / Obrador",
        [
            "Qué resuelve esta plantilla:",
            "▸ El encargo mal anotado es la reclamación más cara de una pastelería: la tarta ya está hecha.",
            "▸ Una ficha por encargo, un registro mensual del dinero y una agenda de entregas por franjas.",
            "",
            "Cómo usar esta plantilla:",
            "▸ Hoja Ficha de Encargo: imprímela y rellénala delante del cliente. Léele en voz alta la "
            "dedicatoria y la fecha de entrega antes de cobrar la señal.",
            "▸ Hoja Registro de Encargos: una línea por encargo. El pendiente de cobro se calcula solo.",
            "▸ Hoja Agenda de Entregas: reparte los encargos por franjas para no acumular recogidas.",
            "▸ Hoja Encargos de Hoy: la que se cuelga cada mañana en el obrador (→ 08 Apertura y Cierre del "
            "Negocio · hoja Apertura del Negocio).",
            "",
            "Alérgenos:",
            "▸ Si el cliente declara una alergia CONFIRMADA, comunícalo por escrito al obrador y comprueba cada "
            "elaboración (→ 12 Control de Alérgenos de Vitrina · hoja Carta de Alérgenos).",
            "",
            "! Protección de datos: esta ficha recoge datos personales y datos de SALUD (las alergias). "
            "Antes de imprimirla, escribe el responsable del tratamiento en la celda verde del pie. Recoge la "
            "casilla de consentimiento explícito del bloque de alérgenos —es distinta de la de comunicaciones "
            "comerciales y ninguna de las dos puede ir marcada de antemano—, guarda las fichas bajo llave, no las "
            "dejes a la vista en el mostrador y destrúyelas al mes de la entrega, salvo lo que necesites "
            "conservar para la facturación.",
            "",
            "Personalización:",
            "▸ Las celdas verdes son editables. Los desplegables de Estado, Forma de pago y Canal se cambian "
            "desde Datos › Validación de datos.",
            "▸ Adapta la política de señal a la tuya antes de imprimir la ficha.",
        ],
    )
    gen11_ficha(wb)
    gen11_registro(wb)
    gen11_agenda(wb)
    gen11_hoy(wb)
    set_metadata(wb, "11")
    path = os.path.join(OUTPUT_DIR, "11-control-encargos.xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════
# 12 — Matriz de Alérgenos de Vitrina (borrador verificable)
# ══════════════════════════════════════════════════════════════════════
# Índices de ALERGENOS: 0 Gluten · 1 Crustáceos · 2 Huevos · 3 Pescado · 4 Cacahuetes
# 5 Soja · 6 Lácteos · 7 Frutos de cáscara · 8 Apio · 9 Mostaza · 10 Sésamo
# 11 Sulfitos · 12 Altramuces · 13 Moluscos
C = "●"   # contiene
T = "○"   # puede contener trazas

MATRIZ = [
    ("Croissant de mantequilla", "Bollería", {0: C, 2: C, 6: C, 5: T, 7: T, 10: T}),
    ("Pain au chocolat", "Bollería", {0: C, 2: C, 6: C, 5: C, 7: T}),
    ("Napolitana de crema", "Bollería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Ensaimada", "Bollería", {0: C, 2: C, 6: T, 5: T, 7: T}),
    ("Brioche", "Bollería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Caracola de pasas", "Bollería", {0: C, 2: C, 6: C, 11: C, 5: T, 7: T}),
    ("Palmera de hojaldre", "Bollería", {0: C, 6: C, 2: T, 5: T, 7: T}),
    ("Croissant relleno de crema", "Bollería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Tarta de queso", "Pastelería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Tarta de manzana", "Pastelería", {0: C, 2: C, 6: C, 11: T, 7: T}),
    ("Milhojas", "Pastelería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Éclair", "Pastelería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Tarta de zanahoria", "Pastelería", {0: C, 2: C, 6: C, 7: C, 5: T}),
    ("Tarta Sacher", "Pastelería", {0: C, 2: C, 6: C, 5: C, 7: T}),
    ("Entremet individual", "Pastelería", {0: C, 2: C, 6: C, 5: C, 7: T}),
    ("Macarons (caja de 6)", "Pastelería", {7: C, 2: C, 6: C, 5: T, 0: T}),
    ("Tartaleta de frutas", "Pastelería", {0: C, 2: C, 6: C, 11: T, 7: T}),
    ("Lionesa o profiterol", "Pastelería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Roscón de Reyes", "Temporada", {0: C, 2: C, 6: C, 7: C, 11: C, 5: T}),
    ("Torrijas", "Temporada", {0: C, 2: C, 6: C, 11: T, 7: T}),
    ("Buñuelos de viento", "Temporada", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Huesos de santo", "Temporada", {7: C, 2: C, 6: T, 0: T}),
    ("Panellets", "Temporada", {7: C, 2: C, 0: T, 6: T}),
    ("Polvorón o mantecado", "Temporada", {0: C, 7: C, 6: T, 10: T, 2: T}),
    ("Turrón de almendra", "Temporada", {7: C, 2: C, 6: T, 5: T, 10: T}),
    ("Mona de Pascua", "Temporada", {0: C, 2: C, 6: C, 5: C, 7: T}),
    ("Brownie", "Bizcochería", {0: C, 2: C, 6: C, 7: C, 5: C}),
    ("Cookie de chocolate", "Bizcochería", {0: C, 2: C, 6: C, 5: C, 7: T}),
    ("Magdalena", "Bizcochería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Bizcocho de yogur", "Bizcochería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Galleta de mantequilla", "Bizcochería", {0: C, 2: C, 6: C, 5: T, 7: T}),
    ("Financier de almendra", "Bizcochería", {7: C, 2: C, 6: C, 0: C, 5: T}),
    ("Quiche", "Salado", {0: C, 2: C, 6: C, 9: T, 8: T, 5: T}),
    ("Empanada de atún", "Salado", {0: C, 3: C, 2: T, 11: T, 8: T}),
    ("Croissant de jamón y queso", "Salado", {0: C, 6: C, 2: C, 9: T, 5: T}),
    ("Pan de molde", "Pan", {0: C, 6: T, 5: T, 10: T}),
    ("Barra", "Pan", {0: C, 10: T, 5: T}),
    ("Hogaza de masa madre", "Pan", {0: C, 10: T, 5: T}),
    ("Chapata", "Pan", {0: C, 10: T, 5: T}),
    ("Focaccia", "Pan", {0: C, 10: T, 5: T, 6: T}),
    ("Crema pastelera", "Elaboración base", {2: C, 6: C, 0: T, 5: T}),
    ("Ganache de chocolate", "Elaboración base", {6: C, 5: C, 7: T, 0: T}),
    ("Nata montada", "Elaboración base", {6: C, 5: T}),
]

# Notas de verificación de las referencias donde el símbolo depende de la marca
# concreta o de la receta de cada casa: se entregan escritas para que el pastelero
# sepa exactamente qué tiene que mirar en la ficha técnica.
NOTAS_MATRIZ = {
    "Caracola de pasas": "Sulfitos: las pasas suelen ir sulfitadas — comprobar mg/kg en la ficha técnica.",
    "Roscón de Reyes": "Sulfitos: comprobar mg/kg en la ficha técnica de la fruta escarchada.",
    "Torrijas": "Sulfitos: si las elaboras con vino, declara el sulfito como ingrediente, no como traza.",
    "Turrón de almendra": "Clara de huevo: verificar receta o variedad (Alicante lleva, Jijona puede no llevar).",
    "Croissant de jamón y queso": "Huevo: sale de la misma masa laminada pincelada que el croissant de mantequilla.",
    "Brownie": "Soja: lecitina de la cobertura de chocolate — comprobar la ficha técnica de tu cobertura.",
    "Cookie de chocolate": "Soja: lecitina de las pepitas de chocolate — comprobar la ficha técnica.",
    "Empanada de atún": "Pescado: solo si es de atún o de bonito. Cambia la referencia si la tuya es de carne.",
    "Polvorón o mantecado": "Sésamo: muchos polvorones llevan ajonjolí como ingrediente — revisa el tuyo.",
    "Macarons (caja de 6)": "Frutos de cáscara: almendra como ingrediente principal, nunca es una traza.",
}

AVISO_BORRADOR = ("PLANTILLA DE PARTIDA — NO LA PUBLIQUES SIN VERIFICARLA. Los símbolos vienen pre-marcados en "
                  "gris como propuesta: contrasta cada referencia con TU receta real y con las fichas técnicas "
                  "de tus proveedores antes de enseñar nada a un cliente.")
LEYENDA = ("●  contiene    ○  puede contener trazas    celda vacía  no contiene    "
           "Los símbolos en gris son la propuesta de partida; al verificarlos, reescríbelos.")
FRASE_ESCRITO = "INFORMACIÓN DE ALÉRGENOS DISPONIBLE POR ESCRITO — SOLICÍTELA A NUESTRO PERSONAL"
# Lista EDITABLE: cada obrador declara los alérgenos que manipula de verdad. Decir
# «los 14» cuando la matriz no recoge ni uno de crustáceos, cacahuetes, altramuces
# o moluscos es un «puede contener» genérico que las propias Instrucciones desaconsejan.
FRASE_CRUZADA = ("En nuestro obrador se manipulan, entre otros, cereales con gluten, huevo, leche, frutos de "
                 "cáscara, soja y sésamo, por lo que puede haber contaminación cruzada. Esta advertencia "
                 "complementa la información por escrito: no la sustituye.")


def gen12_matriz(wb):
    ws = wb.create_sheet("Matriz Alérgenos")
    ws.sheet_properties.tabColor = "FFD700"
    cols = {"A": 28, "B": 17}
    for i in range(3, 17):
        cols[get_column_letter(i)] = 5
    cols.update({"Q": 9, "R": 26, "S": 22, "T": 13, "U": 15, "V": 52})
    widths(ws, cols)

    sheet_title(ws, 1, 22, "Matriz de Alérgenos de Vitrina — Borrador de Trabajo")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=22)
    c = ws.cell(row=2, column=1, value=AVISO_BORRADOR)
    c.font = alert_font
    c.fill = alert_fill
    c.alignment = left_align
    c.border = thin_border
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=22)
    c = ws.cell(row=3, column=1, value=LEYENDA)
    c.font = note_font
    c.alignment = left_align

    headers = ["Producto", "Categoría"] + [a[0] for a in ALERGENOS] + [
        "Total ●", "Proveedor / Nº de ficha técnica",
        "Verificado con ficha técnica (fecha)", "Firma", "Estado", "Notas de verificación"]
    table_header(ws, 4, headers, rotate=set(range(3, 17)), height=95)

    dv_mark = DataValidation(type="list", formula1='"●,○"', allow_blank=True)
    ws.add_data_validation(dv_mark)

    row = 5
    first = row
    for prod, cat, marcas in MATRIZ:
        paint(ws, row, 1, prod)
        paint(ws, row, 2, cat, font=Font(name="Calibri", size=10, color="666666"), align=center_align)
        for i in range(14):
            paint(ws, row, 3 + i, marcas.get(i), font=gray_mark_font, align=center_align)
        paint(ws, row, 17, f'=COUNTIF(C{row}:P{row},"●")', font=bold_font, align=center_align, fmt="0")
        paint(ws, row, 18, None, fill=input_fill)
        paint(ws, row, 19, None, fill=input_fill, align=center_align)
        paint(ws, row, 20, None, fill=input_fill, align=center_align)
        paint(ws, row, 21, f'=IF($S{row}="","NO PUBLICAR","Verificado")', font=bold_font,
              align=center_align)
        paint(ws, row, 22, NOTAS_MATRIZ.get(prod), font=note_font, fill=input_fill)
        row += 1
    for _ in range(4):
        for col in range(1, 23):
            if col == 17:
                paint(ws, row, col, f'=COUNTIF(C{row}:P{row},"●")', align=center_align, fmt="0")
            elif col == 21:
                paint(ws, row, col, f'=IF($S{row}="","NO PUBLICAR","Verificado")', align=center_align)
            else:
                paint(ws, row, col, None, fill=input_fill,
                      align=center_align if 1 < col < 22 else left_align)
        row += 1
    last = row - 1
    dv_mark.add(f"C{first}:P{last}")

    ws.conditional_formatting.add(
        f"U{first}:U{last}",
        CellIsRule(operator="equal", formula=['"NO PUBLICAR"'], fill=red_fill,
                   font=Font(name="Calibri", size=11, bold=True, color="B71C1C")))
    ws.conditional_formatting.add(
        f"U{first}:U{last}",
        CellIsRule(operator="equal", formula=['"Verificado"'], fill=green_ok_fill))

    row += 1
    paint(ws, row, 1, "Nº de referencias que lo CONTIENEN (●)", font=bold_font, fill=gold_fill)
    paint(ws, row, 2, None, fill=gold_fill)
    for i in range(3, 17):
        letter = get_column_letter(i)
        paint(ws, row, i, f'=COUNTIF({letter}{first}:{letter}{last},"●")',
              font=bold_font, fill=gold_fill, align=center_align, fmt="0")
    for i in range(17, 23):
        paint(ws, row, i, None, fill=gold_fill)
    row += 1
    paint(ws, row, 1, "Nº de referencias con TRAZAS (○)", font=bold_font, fill=gray_fill)
    paint(ws, row, 2, None, fill=gray_fill)
    for i in range(3, 17):
        letter = get_column_letter(i)
        paint(ws, row, i, f'=COUNTIF({letter}{first}:{letter}{last},"○")',
              font=bold_font, fill=gray_fill, align=center_align, fmt="0")
    for i in range(17, 23):
        paint(ws, row, i, None, fill=gray_fill)
    row += 1
    paint(ws, row, 1, "Referencias pendientes de verificar", font=bold_font)
    paint(ws, row, 2, f'=COUNTIF(U{first}:U{last},"NO PUBLICAR")', font=bold_font,
          align=center_align, fmt="0")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=22)
    paint(ws, row, 3, "Mientras este número no sea cero, la carta de alérgenos no está lista para enseñarse.",
          font=note_font)
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)
    return first, last


def gen12_carta(wb, first, last):
    ws = wb.create_sheet("Carta de Alérgenos")
    ws.sheet_properties.tabColor = "FF8C00"
    cols = {"A": 30}
    for i in range(2, 16):
        cols[get_column_letter(i)] = 5
    widths(ws, cols)
    ws.sheet_view.showGridLines = False

    sheet_title(ws, 1, 15, "Carta de Alérgenos",
                "Establecimiento: ______________________________     "
                "Fecha de revisión: ____/____/________")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=15)
    c = ws.cell(row=3, column=1,
                value=("Documento para enseñar al cliente. Se rellena solo desde la hoja Matriz Alérgenos: "
                       "verifica allí cada referencia antes de imprimir esta carta.   "
                       "●  contiene    ○  puede contener trazas"))
    c.font = note_font
    c.alignment = left_align

    table_header(ws, 5, ["Producto"] + [a[0] for a in ALERGENOS],
                 rotate=set(range(2, 16)), height=95)

    row = 6
    src = "'Matriz Alérgenos'"
    for i, srcrow in enumerate(range(first, last + 1)):
        paint(ws, row, 1, f'=IF({src}!A{srcrow}="","",{src}!A{srcrow})')
        for j in range(14):
            letter = get_column_letter(3 + j)
            paint(ws, row, 2 + j,
                  f'=IF({src}!{letter}{srcrow}="","",{src}!{letter}{srcrow})',
                  font=Font(name="Calibri", size=12), align=center_align)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=15)
    c = ws.cell(row=row, column=1, value=FRASE_CRUZADA)
    c.font = note_font
    c.alignment = left_align
    fill_range(ws, row, 1, row + 1, 15, input_fill, thin_border)
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    ws.cell(row=row, column=1,
            value="Revisado por: _________________________     Firma: _________________________").font = bold_font
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(5, 5), landscape=True, gridlines=False)


def gen12_cartel(wb):
    ws = wb.create_sheet("Cartel Vitrina")
    ws.sheet_properties.tabColor = "8B4513"
    widths(ws, {"A": 2, "B": 24, "C": 56, "D": 30})
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:D2")
    c = ws.cell(row=2, column=2, value="Información sobre Alérgenos")
    c.font = Font(name="Calibri", size=22, bold=True, color="1A1A1A")
    c.alignment = center_align
    ws.row_dimensions[2].height = 34

    ws.merge_cells("B4:D5")
    c = ws.cell(row=4, column=2, value=FRASE_ESCRITO)
    c.font = Font(name="Calibri", size=14, bold=True, color="1A1A1A")
    c.alignment = center_align
    fill_range(ws, 4, 2, 5, 4, gold_fill, thin_border)
    ws.row_dimensions[4].height = 26

    ws.merge_cells("B7:D8")
    c = ws.cell(row=7, column=2, value=FRASE_CRUZADA)
    c.font = Font(name="Calibri", size=11, color="333333")
    c.alignment = center_align
    fill_range(ws, 7, 2, 8, 4, input_fill, thin_border)

    ws.merge_cells("B10:D10")
    c = ws.cell(row=10, column=2, value="Los 14 alérgenos de declaración obligatoria")
    c.font = section_font
    c.alignment = center_align
    fill_range(ws, 10, 2, 10, 4, header_fill)
    ws.row_dimensions[10].height = 22

    row = 11
    for col, h in ((2, "Alérgeno"), (3, "Qué incluye"), (4, "Nuestros productos que lo contienen")):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    for nombre, desc in ALERGENOS:
        paint(ws, row, 2, nombre, font=bold_font)
        paint(ws, row, 3, desc, font=Font(name="Calibri", size=10, color="333333"))
        paint(ws, row, 4, None, fill=input_fill)
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2,
                value="Establecimiento: ______________________________          "
                      "Última revisión: ____/____/________")
    c.font = bold_font
    c.alignment = center_align
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2,
                value="Reglamento (UE) 1169/2011 y Real Decreto 126/2015, sobre la información alimentaria de "
                      "los alimentos que se presentan sin envasar.")
    c.font = small_font
    c.alignment = center_align
    row += 2
    footer_line(ws, row, col=2)
    print_setup(ws, header_rows=None, landscape=False, one_page=True,
                gridlines=False, print_area=f"A1:D{row}")


def gen12_etiquetas(wb):
    ws = wb.create_sheet("Etiquetas Vitrina")
    ws.sheet_properties.tabColor = "FFD700"
    widths(ws, {"A": 2, "B": 30, "C": 2, "D": 30, "E": 2, "F": 30})
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:F1")
    c = ws.cell(row=1, column=2, value="Etiquetas de Vitrina")
    c.font = title_font
    c.alignment = left_align
    ws.merge_cells("B2:F2")
    c = ws.cell(row=2, column=2,
                value="Rellena, imprime en cartulina y recorta por el borde. Abrevia los alérgenos: "
                      "GLU gluten · HUE huevos · LAC lácteos · FCA frutos de cáscara · SOJ soja · SES sésamo · SUL sulfitos.")
    c.font = note_font
    c.alignment = left_align

    row = 4
    for _ in range(6):
        for col in (2, 4, 6):
            paint(ws, row, col, None, font=Font(name="Calibri", size=13, bold=True),
                  fill=input_fill, align=center_align, border=False).border = cut_border
            paint(ws, row + 1, col, None, font=Font(name="Calibri", size=12),
                  fill=input_fill, align=center_align, border=False).border = cut_border
            paint(ws, row + 2, col, None, font=Font(name="Calibri", size=9, color="555555"),
                  fill=input_fill, align=center_align, border=False).border = cut_border
        ws.row_dimensions[row].height = 26
        ws.row_dimensions[row + 1].height = 20
        ws.row_dimensions[row + 2].height = 18
        for col in (2, 4, 6):
            ws.cell(row=row, column=col).value = None
        row += 4

    guia = row + 1
    ws.merge_cells(start_row=guia, start_column=2, end_row=guia, end_column=6)
    ws.cell(row=guia, column=2,
            value="Cada etiqueta: primera línea el producto, segunda el precio por unidad, tercera los alérgenos "
                  "(→ hoja Carta de Alérgenos).").font = note_font
    footer_line(ws, guia + 2, col=2)
    print_setup(ws, header_rows=None, landscape=False, one_page=True,
                gridlines=False, print_area=f"A1:F{guia + 2}")


def generate_12():
    wb = Workbook()
    instructions_sheet(
        wb,
        "12 — Control de Alérgenos de Vitrina · 14 Alérgenos de Declaración Obligatoria",
        [
            "! Esta plantilla se entrega como BORRADOR. Los símbolos vienen pre-marcados en gris como propuesta "
            "de partida. No publiques ni enseñes nada sin contrastar cada referencia con tu receta real y con "
            "las fichas técnicas de tus proveedores. La responsabilidad de la información es del establecimiento.",
            "",
            "Qué exige la ley:",
            "▸ El Reglamento (UE) 1169/2011 y el Real Decreto 126/2015 obligan a informar de los 14 alérgenos "
            "también en los alimentos que se venden sin envasar: vitrina, granel, obrador y cafetería.",
            "▸ En producto sin envasar la información tiene que estar disponible POR ESCRITO y ser accesible "
            "antes de la compra. La vía oral solo vale si hay un cartel visible que indique que la información "
            "escrita se puede solicitar.",
            "▸ Un «puede contener» genérico no sustituye a declarar los alérgenos que llevas de forma intencionada.",
            "",
            "Cómo usar esta plantilla:",
            "▸ Hoja Matriz Alérgenos: repasa referencia por referencia. Marca ● si el producto lo contiene y "
            "○ si puede contener trazas por contaminación cruzada en tu obrador.",
            "▸ Anota el proveedor y el número de ficha técnica que respalda cada decisión, pon la fecha de "
            "verificación y firma. Hasta que no pongas la fecha, la columna Estado dice NO PUBLICAR.",
            "▸ La columna Notas de verificación te dice qué mirar en las referencias donde el símbolo depende "
            "de la marca o de la receta: pasas y fruta escarchada (sulfitos), turrón (clara de huevo), "
            "coberturas de chocolate (lecitina de soja), polvorones (ajonjolí).",
            "▸ La frase de contaminación cruzada del cartel y de la carta está en celda verde: sustituye la "
            "lista por los alérgenos que se manipulan de verdad en tu obrador. Declarar «los 14» cuando no "
            "trabajas con crustáceos ni con moluscos es el «puede contener» genérico que la ley desaconseja.",
            "▸ Hoja Carta de Alérgenos: se rellena sola desde la matriz. Es el documento que se enseña al "
            "cliente y el que pide el inspector. Imprímela cuando no quede ninguna referencia sin verificar.",
            "▸ Hoja Cartel Vitrina: cuélgalo a la vista del cliente, junto a la vitrina.",
            "▸ Hoja Etiquetas Vitrina: una etiqueta por referencia, con precio y alérgenos abreviados.",
            "",
            "Cuándo hay que revisarla:",
            "▸ Siempre que cambies una receta, un ingrediente o un proveedor.",
            "▸ Al menos una vez al año, aunque no haya cambiado nada, y siempre antes de una campaña.",
            "",
            "Personalización:",
            "▸ Las celdas verdes son editables. Los símbolos ● y ○ se eligen del desplegable de cada casilla.",
            "▸ Hay cuatro filas libres al final de la matriz para tus referencias propias.",
            "▸ Referencias del kit: las etiquetas de vitrina se colocan en la apertura "
            "(→ 08 Apertura y Cierre del Negocio · hoja Apertura del Negocio) y los alérgenos declarados por "
            "el cliente se anotan en la ficha de encargo (→ 11 Control de Encargos).",
        ],
    )
    first, last = gen12_matriz(wb)
    gen12_carta(wb, first, last)
    gen12_cartel(wb)
    gen12_etiquetas(wb)
    set_metadata(wb, "12")
    path = os.path.join(OUTPUT_DIR, "12-control-alergenos-vitrina.xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════
# 13 — Registro de Temperaturas + Recepción + Etiquetas de Elaborado
# ══════════════════════════════════════════════════════════════════════
EQUIPOS = [
    ("Cámara de refrigeración", "0 a 4 °C", ("Mañana", "Tarde"), ("between", 0, 4)),
    ("Cámara de masas", "2 a 6 °C", ("Mañana", "Tarde"), ("between", 2, 6)),
    ("Congelador", "≤ −18 °C", ("Mañana", "Tarde"), ("greaterThan", -18, None)),
    ("Vitrina 1", "2 a 6 °C", ("Mañana", "Tarde"), ("between", 2, 6)),
    ("Vitrina 2", "2 a 6 °C", ("Mañana", "Tarde"), ("between", 2, 6)),
    ("Abatidor", "+65 a +10 °C en menos de 2 h", ("Tª final", "Minutos"), ("abatidor", None, None)),
]

VIDAS_UTILES = [
    ("Crema pastelera", "0 a 4 °C, tapada en contacto", "48 a 72 h", "Enfriar en placa fina, nunca en el cazo"),
    ("Nata montada", "0 a 4 °C", "24 h", "No se recongela"),
    ("Ganache de chocolate", "0 a 4 °C", "5 a 7 días", "Atemperar antes de usar"),
    ("Mousse o bavarois", "0 a 4 °C", "3 días", "Montado y sin acabar"),
    ("Merengue italiano", "0 a 4 °C", "3 días", "Tapado, sin humedad"),
    ("Trufa cocida", "0 a 4 °C", "5 días", ""),
    ("Curd de limón", "0 a 4 °C", "5 a 7 días", "En bote esterilizado"),
    ("Almíbar o jarabe", "0 a 4 °C", "15 días", "Grado de azúcar alto: más duración"),
    ("Fruta cortada", "0 a 4 °C", "24 h", "Con protección de gelatina o brillo"),
    ("Masa laminada cruda", "≤ −18 °C", "1 mes", "Congelar formada y filmada"),
    ("Bizcocho base", "≤ −18 °C", "3 meses", "Filmado por unidades"),
    ("Tarta montada con nata", "0 a 4 °C", "24 a 48 h", "Vitrina a 2-6 °C"),
    ("Bollería cocida", "Ambiente, en vitrina", "El mismo día", "Al día siguiente, a merma o a promoción"),
    ("Pan cocido", "Ambiente", "24 a 48 h", "Según formato y corteza"),
]

CRITERIOS_RECHAZO = (
    "Rechaza la entrega si: el producto refrigerado llega por encima de 4 °C · el congelado presenta signos de "
    "descongelación (escarcha, bloque deformado, líquido en la caja) · el envase está roto, abierto o mojado · "
    "la fecha de caducidad o de consumo preferente está a menos de 2 días · el vehículo llega sucio o sin frío · "
    "el albarán no coincide con lo que recibes. Anota el motivo, haz una foto y avisa al proveedor el mismo día. "
    "Aceptado con reserva: la mercancía se recibe, pero la incidencia se anota en el albarán antes de firmarlo "
    "y se fotografía; el producto se aparta identificado y no se usa hasta que el proveedor conteste."
)


def gen13_temperaturas(wb):
    ws = wb.create_sheet("Registro de Temperaturas")
    ws.sheet_properties.tabColor = "FFD700"
    widths(ws, {"A": 6})
    for i in range(2, 14):
        widths(ws, {get_column_letter(i): 9})
    widths(ws, {"N": 42, "O": 14})

    sheet_title(ws, 1, 15, "Registro de Temperaturas",
                "Mes: ________________     Año: ________     Establecimiento: ______________________________")

    # cabecera de equipos (fila 4) + tomas (fila 5)
    col = 2
    for nombre, rango, tomas, _ in EQUIPOS:
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        c = ws.cell(row=4, column=col, value=f"{nombre}\n({rango})")
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        for k in (col, col + 1):
            ws.cell(row=4, column=k).border = thin_border
        for j, toma in enumerate(tomas):
            cc = ws.cell(row=5, column=col + j, value=toma)
            cc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cc.fill = header_fill
            cc.alignment = center_align
            cc.border = thin_border
        col += 2
    for c0, label in ((1, "Día"), (14, "Acción correctiva"), (15, "Firma")):
        ws.merge_cells(start_row=4, start_column=c0, end_row=5, end_column=c0)
        cell = ws.cell(row=4, column=c0, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(row=5, column=c0).border = thin_border
    ws.row_dimensions[4].height = 42

    row = 6
    first = row
    for dia in range(1, 32):
        paint(ws, row, 1, dia, font=Font(name="Calibri", size=10, color="666666"),
              align=center_align, fmt="0")
        for i in range(2, 14):
            paint(ws, row, i, None, fill=input_fill, align=center_align, fmt="0.0")
        paint(ws, row, 14, None, fill=input_fill)
        paint(ws, row, 15, None, fill=input_fill, align=center_align)
        row += 1
    last = row - 1

    # Reglas de FÓRMULA con guarda de vacío. Con las reglas «entre valores» de
    # Excel, la celda vacía se evalúa como 0: la hoja se abría con las 248
    # casillas del mes en rojo antes de anotar una sola temperatura.
    col = 2
    for nombre, rango, tomas, regla in EQUIPOS:
        kind, a, b = regla
        c1, c2 = get_column_letter(col), get_column_letter(col + 1)
        if kind == "between":
            ws.conditional_formatting.add(
                f"{c1}{first}:{c2}{last}",
                FormulaRule(formula=[f'AND({c1}{first}<>"",OR({c1}{first}<{a},{c1}{first}>{b}))'],
                            fill=red_fill))
        elif kind == "greaterThan":
            ws.conditional_formatting.add(
                f"{c1}{first}:{c2}{last}",
                FormulaRule(formula=[f'AND({c1}{first}<>"",{c1}{first}>{a})'], fill=red_fill))
        else:  # abatidor: Tª final ≤ 10 °C y ciclo < 120 min
            ws.conditional_formatting.add(
                f"{c1}{first}:{c1}{last}",
                FormulaRule(formula=[f'AND({c1}{first}<>"",{c1}{first}>10)'], fill=red_fill))
            ws.conditional_formatting.add(
                f"{c2}{first}:{c2}{last}",
                FormulaRule(formula=[f'AND({c2}{first}<>"",{c2}{first}>120)'], fill=red_fill))
        col += 2

    row += 1
    paint(ws, row, 1, "Tomas", font=bold_font, fill=gold_fill)
    paint(ws, row, 2, f"=COUNT(B{first}:M{last})", font=bold_font, fill=gold_fill,
          align=center_align, fmt="0")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=15)
    paint(ws, row, 3,
          "Número de lecturas anotadas este mes. Toda casilla en rojo exige una acción correctiva escrita.",
          font=note_font, fill=gold_fill)
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=15)
    c = ws.cell(row=row, column=1,
                value=("Qué hacer si una temperatura se sale de rango: comprueba que la puerta cierra y que el "
                       "equipo no está sobrecargado, traslada el género a otro equipo, anota la acción en esta "
                       "misma fila y avisa al servicio técnico. Si el producto ha estado fuera de rango más de "
                       "dos horas, valóralo antes de venderlo y regístralo como merma "
                       "(→ 10 Plan de Producción Semanal · hoja Producido vs Vendido)."))
    c.font = note_font
    c.alignment = left_align
    c.fill = amber_fill
    c.border = thin_border
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    ws.cell(row=row, column=1,
            value="Revisión semanal del responsable — Semana 1: __________   Semana 2: __________   "
                  "Semana 3: __________   Semana 4: __________").font = bold_font
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 5), landscape=True)


RECEPCION_FILAS = 30


def gen13_recepcion(wb):
    ws = wb.create_sheet("Recepción de Mercancía")
    ws.sheet_properties.tabColor = "FF8C00"
    widths(ws, {"A": 12, "B": 24, "C": 14, "D": 30, "E": 16, "F": 13, "G": 12,
                "H": 20, "I": 14, "J": 30, "K": 14})
    sheet_title(ws, 1, 11, "Recepción de Mercancía",
                "Mes: ________________     Año: ________     "
                "Una línea por entrega. Comprueba la temperatura ANTES de firmar el albarán.")
    table_header(ws, 4, ["Fecha", "Proveedor", "Nº albarán", "Producto", "Lote", "Caducidad",
                         "Tª entrega (°C)", "Vehículo y embalaje", "Aceptado",
                         "Motivo del rechazo", "Firma"])

    dv_estado = DataValidation(type="list", formula1='"Correcto,Deficiente"', allow_blank=True)
    dv_acepta = DataValidation(type="list", formula1='"Aceptado,Rechazado,Aceptado con reserva"',
                               allow_blank=True)
    ws.add_data_validation(dv_estado)
    ws.add_data_validation(dv_acepta)

    row = 5
    first = row
    for _ in range(RECEPCION_FILAS):
        paint(ws, row, 1, None, fill=input_fill, align=center_align, fmt="dd/mm/yyyy")
        paint(ws, row, 2, None, fill=input_fill)
        paint(ws, row, 3, None, fill=input_fill, align=center_align)
        paint(ws, row, 4, None, fill=input_fill)
        paint(ws, row, 5, None, fill=input_fill, align=center_align)
        paint(ws, row, 6, None, fill=input_fill, align=center_align, fmt="dd/mm/yyyy")
        paint(ws, row, 7, None, fill=input_fill, align=center_align, fmt="0.0")
        paint(ws, row, 8, None, fill=input_fill, align=center_align)
        paint(ws, row, 9, None, fill=input_fill, align=center_align)
        paint(ws, row, 10, None, fill=input_fill)
        paint(ws, row, 11, None, fill=input_fill, align=center_align)
        ws.row_dimensions[row].height = 20
        row += 1
    last = row - 1
    dv_estado.add(f"H{first}:H{last}")
    dv_acepta.add(f"I{first}:I{last}")

    ws.conditional_formatting.add(
        f"I{first}:I{last}",
        CellIsRule(operator="equal", formula=['"Rechazado"'], fill=red_fill))
    ws.conditional_formatting.add(
        f"I{first}:I{last}",
        CellIsRule(operator="equal", formula=['"Aceptado"'], fill=green_ok_fill))
    ws.conditional_formatting.add(
        f"G{first}:G{last}",
        CellIsRule(operator="greaterThan", formula=["4"], fill=amber_fill))

    paint(ws, row, 1, "Entregas anotadas", font=bold_font, fill=gold_fill)
    paint(ws, row, 2, f'=COUNTIF(B{first}:B{last},"?*")', font=bold_font, fill=gold_fill,
          align=center_align, fmt="0")
    paint(ws, row, 3, "Rechazadas", font=bold_font, fill=gold_fill)
    paint(ws, row, 4, f'=COUNTIF(I{first}:I{last},"Rechazado")', font=bold_font, fill=gold_fill,
          align=center_align, fmt="0")
    for col in range(5, 12):
        paint(ws, row, col, None, fill=gold_fill)
    row += 2

    block_row(ws, row, 11, "  Criterios de Rechazo")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=11)
    c = ws.cell(row=row, column=1, value=CRITERIOS_RECHAZO)
    c.font = note_font
    c.alignment = left_align
    c.fill = amber_fill
    c.border = thin_border
    row += 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    ws.cell(row=row, column=1,
            value="La temperatura de entrega se mide con termómetro de sonda entre dos envases, nunca clavándolo "
                  "en el producto.").font = note_font
    row += 2
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=True)


def gen13_etiquetas(wb):
    ws = wb.create_sheet("Etiquetas de Elaborado")
    ws.sheet_properties.tabColor = "8B4513"
    widths(ws, {"A": 2, "B": 30, "C": 2, "D": 30, "E": 2, "F": 30})
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:F1")
    c = ws.cell(row=1, column=2, value="Etiquetas de Elaborado")
    c.font = title_font
    c.alignment = left_align
    ws.merge_cells("B2:F2")
    c = ws.cell(row=2, column=2,
                value="Toda elaboración que entre en cámara sale etiquetada. Sin etiqueta no hay trazabilidad: "
                      "ni sabes qué es, ni hasta cuándo sirve. Vidas útiles orientativas en la hoja Vidas Útiles.")
    c.font = note_font
    c.alignment = left_align

    etiquetas = ["Producto", "Fecha de elaboración", "Fecha límite de consumo", "Lote", "Responsable"]
    row = 4
    for _ in range(8):
        for i, label in enumerate(etiquetas):
            for col in (2, 4, 6):
                cell = ws.cell(row=row + i, column=col)
                cell.value = f"{label}:" if label else None
                cell.font = Font(name="Calibri", size=9, bold=(i == 0), color="555555")
                cell.alignment = left_align
                cell.fill = input_fill
                cell.border = cut_border
            ws.row_dimensions[row + i].height = 17
        row += len(etiquetas) + 1
    print_setup(ws, header_rows=None, landscape=False, one_page=True,
                gridlines=False, print_area=f"A1:F{row - 1}")


def gen13_vidas(wb):
    ws = wb.create_sheet("Vidas Útiles")
    ws.sheet_properties.tabColor = "FFD700"
    widths(ws, {"A": 30, "B": 30, "C": 22, "D": 44})
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 4, "Vidas Útiles Orientativas por Elaboración",
                "Cuélgala en el obrador. Cada obrador fija las suyas: estas son un punto de partida.")
    table_header(ws, 4, ["Elaboración", "Conservación", "Vida útil orientativa", "Notas"])
    row = 5
    for elab, cons, vida, nota in VIDAS_UTILES:
        paint(ws, row, 1, elab, font=bold_font)
        paint(ws, row, 2, cons)
        paint(ws, row, 3, vida, align=center_align)
        paint(ws, row, 4, nota, font=note_font)
        ws.row_dimensions[row].height = 22
        row += 1
    for _ in range(4):
        for col in range(1, 5):
            paint(ws, row, col, None, fill=input_fill)
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=4)
    c = ws.cell(row=row, column=1,
                value=("Estas vidas útiles son ORIENTATIVAS y suponen elaboración correcta, enfriado rápido, "
                       "envase limpio y cadena de frío sin cortes. La vida útil de tus elaboraciones la fijas tú "
                       "y debe constar en tu plan de APPCC; si tienes dudas con una elaboración concreta, "
                       "consulta con tu asesor en seguridad alimentaria o encarga un estudio de vida útil."))
    c.font = note_font
    c.alignment = left_align
    c.fill = amber_fill
    c.border = thin_border
    row += 4
    footer_line(ws, row)
    print_setup(ws, header_rows=(4, 4), landscape=False, gridlines=False)


def generate_13():
    wb = Workbook()
    instructions_sheet(
        wb,
        "13 — Registro de Temperaturas y Recepción de Mercancía",
        [
            "Qué resuelve esta plantilla:",
            "▸ El kit te pide registrar temperaturas en varios checklists: esta es la hoja donde se registran.",
            "▸ Registro de temperaturas, recepción de mercancía y etiquetado de elaborados son los tres "
            "registros que se piden en una inspección de un obrador.",
            "",
            "Cómo usar esta plantilla:",
            "▸ Hoja Registro de Temperaturas: una hoja por mes. Dos tomas al día, mañana y tarde. Las lecturas "
            "fuera del rango objetivo se marcan solas en rojo: escribe siempre la acción correctiva.",
            "▸ Imprime una hoja por mes y guárdala firmada. Los registros son la prueba de que el control existe.",
            "▸ Hoja Recepción de Mercancía: comprueba la temperatura antes de firmar el albarán. Si rechazas "
            "algo, anota el motivo el mismo día.",
            "▸ Hoja Etiquetas de Elaborado: rellena, imprime y pega. Toda elaboración que entre en cámara "
            "sale etiquetada.",
            "▸ Hoja Vidas Útiles: la referencia de cuánto dura cada elaboración. Es orientativa.",
            "",
            "Rangos objetivo de partida:",
            "▸ Cámara de refrigeración 0 a 4 °C · cámara de masas 2 a 6 °C · congelador igual o por debajo de "
            "−18 °C · vitrinas 2 a 6 °C · abatidor de +65 a +10 °C en menos de 2 horas.",
            "▸ Ajusta los rangos a los de tu plan de APPCC si son distintos: se cambian en Inicio › Formato "
            "condicional › Administrar reglas.",
            "",
            "Personalización:",
            "▸ Las celdas verdes son editables. Cambia los nombres de los equipos por los tuyos.",
            "▸ Referencias del kit: las tomas de vitrina salen de la apertura y del cierre "
            "(→ 08 Apertura y Cierre del Negocio) y las de cámaras, del obrador "
            "(→ 01 Apertura y Cierre · hojas Apertura Obrador y Cierre Obrador).",
        ],
    )
    gen13_temperaturas(wb)
    gen13_recepcion(wb)
    gen13_etiquetas(wb)
    gen13_vidas(wb)
    set_metadata(wb, "13")
    path = os.path.join(OUTPUT_DIR, "13-registro-temperaturas-recepcion.xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════
# Verificación
# ══════════════════════════════════════════════════════════════════════
NO_LATINO = re.compile(
    "[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff\uac00-\ud7af\u0400-\u04ff"
    "\u0600-\u06ff\u0590-\u05ff\u0e00-\u0e7f]")


def verify(path):
    logging.disable(logging.CRITICAL)
    from pycel import ExcelCompiler

    wbf = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    xl = ExcelCompiler(filename=path)

    formulas = 0
    sin_cache = []
    vacias_por_diseno = 0
    errores_pycel = []
    no_latino = 0
    for wsf, wsv in zip(wbf.worksheets, wbv.worksheets):
        for row in wsf.iter_rows():
            for c in row:
                if isinstance(c.value, str) and NO_LATINO.search(c.value):
                    no_latino += 1
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
                    cached = wsv[c.coordinate].value
                    if cached is None:
                        try:
                            with open(os.devnull, "w") as dn, contextlib.redirect_stderr(dn):
                                v = xl.evaluate(f"'{wsf.title}'!{c.coordinate}")
                        except Exception as e:  # noqa: BLE001
                            errores_pycel.append(f"{wsf.title}!{c.coordinate}: {type(e).__name__}")
                            continue
                        if v == "" or v is None:
                            vacias_por_diseno += 1
                        else:
                            sin_cache.append(f"{wsf.title}!{c.coordinate}")
    return {
        "hojas": len(wbf.worksheets),
        "formulas": formulas,
        "sin_cache": sin_cache,
        "formulas_vacias_por_diseno": vacias_por_diseno,
        "errores_pycel": errores_pycel,
        "no_latino": no_latino,
        "creator": wbf.properties.creator,
        "subject": wbf.properties.subject,
        "bio_vieja": any(
            isinstance(c.value, str) and ("29 años" in c.value or "15 años de consultoría" in c.value)
            for ws in wbf.worksheets for row in ws.iter_rows() for c in row),
        "chefbusiness": any(
            isinstance(c.value, str) and "chefbusiness" in c.value.lower()
            for ws in wbf.worksheets for row in ws.iter_rows() for c in row),
    }


def main():
    paths = [generate_08(), generate_10(), generate_11(), generate_12(), generate_13()]
    for p in paths:
        print(f"  generado  {os.path.basename(p)}")

    print("\nInyectando cache de valores...")
    subprocess.run([sys.executable, INJECT] + paths, check=True)

    print("\nVerificación:")
    ok = True
    for p in paths:
        v = verify(p)
        good = (not v["sin_cache"] and not v["errores_pycel"] and v["no_latino"] == 0
                and v["creator"] == "AI Chef Pro" and not v["bio_vieja"] and not v["chefbusiness"])
        ok &= good
        print(f"  {'OK  ' if good else 'FALLA'} {os.path.basename(p)}: "
              f"hojas={v['hojas']} formulas={v['formulas']} "
              f"sin_cache={len(v['sin_cache'])} vacias_por_diseno={v['formulas_vacias_por_diseno']} "
              f"errores_pycel={len(v['errores_pycel'])} no_latino={v['no_latino']} "
              f"creator={v['creator']}")
        if v["sin_cache"]:
            print("        sin cache:", v["sin_cache"][:10])
        if v["errores_pycel"]:
            print("        errores pycel:", v["errores_pycel"][:10])
    print("\nTODO OK" if ok else "\nHAY FALLOS")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
