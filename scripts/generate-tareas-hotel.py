#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Hotel (15 archivos, 46 plantillas, 11 departamentos).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-hotel"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Department colors
BUFFET_COLOR = "FFF3E0"
ROOMSERVICE_COLOR = "E3F2FD"
BANQUETE_COLOR = "F3E5F5"
DESAYUNO_COLOR = "E8F5E9"
MINIBAR_COLOR = "FCE4EC"
ADMIN_COLOR = "FFF8E1"
EVENTO_COLOR = "E0F2F1"
LIMPIEZA_COLOR = "EFEBE9"
RECEPCION_COLOR = "E8EAF6"
HOUSEKEEPING_COLOR = "F3E5F5"
PISCINA_COLOR = "E0F7FA"
TERRAZA_COLOR = "F1F8E9"
MANTENIMIENTO_COLOR = "FFF3E0"
ADMIN_OFICINA_COLOR = "FFF8E1"
SPA_COLOR = "FCE4EC"
SEGURIDAD_COLOR = "ECEFF1"
AREAS_PUBLICAS_COLOR = "F5F5F5"

ZONE_COLORS = {
    "Buffet": BUFFET_COLOR,
    "Room Service": ROOMSERVICE_COLOR,
    "Banquetes": BANQUETE_COLOR,
    "Desayuno": DESAYUNO_COLOR,
    "Minibar": MINIBAR_COLOR,
    "Admin": ADMIN_COLOR,
    "Evento": EVENTO_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Cocina": BUFFET_COLOR,
    "Sala": BANQUETE_COLOR,
    "General": LIGHT_GRAY,
    "Recepción": RECEPCION_COLOR,
    "Conserjería": RECEPCION_COLOR,
    "Guest Exp.": RECEPCION_COLOR,
    "Housekeeping": HOUSEKEEPING_COLOR,
    "Lencería": HOUSEKEEPING_COLOR,
    "Piscina": PISCINA_COLOR,
    "Terraza": TERRAZA_COLOR,
    "Mantenimiento": MANTENIMIENTO_COLOR,
    "HVAC": MANTENIMIENTO_COLOR,
    "Fontanería": MANTENIMIENTO_COLOR,
    "Electricidad": MANTENIMIENTO_COLOR,
    "Revenue": ADMIN_OFICINA_COLOR,
    "Reservas": ADMIN_OFICINA_COLOR,
    "Contabilidad": ADMIN_OFICINA_COLOR,
    "RRHH": ADMIN_OFICINA_COLOR,
    "Spa": SPA_COLOR,
    "Vestuarios": SPA_COLOR,
    "Seguridad": SEGURIDAD_COLOR,
    "Áreas Públicas": AREAS_PUBLICAS_COLOR,
    "Lobby": AREAS_PUBLICAS_COLOR,
    "Pasillos": AREAS_PUBLICAS_COLOR,
    "Baños": AREAS_PUBLICAS_COLOR,
    "Parking": AREAS_PUBLICAS_COLOR,
    "Pool Bar": PISCINA_COLOR,
    "Lobby Bar": EVENTO_COLOR,
    "Snack Bar": BUFFET_COLOR,
    "À la Carte": BANQUETE_COLOR,
}

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Hotel Integral"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for section_title, tasks in sections:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for idx, (task, zone) in enumerate(tasks, 1):
            zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")

            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=zone).font = data_font
            ws.cell(row=row, column=3).fill = zone_fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align

            check_cell = ws.cell(row=row, column=5)
            check_cell.font = checkbox_font
            check_cell.alignment = center_align
            check_cell.border = thin_border
            dv.add(check_cell)

            ws.cell(row=row, column=6).fill = input_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


# ═══════════════════════════════════════════════════════════
# 01 — DESAYUNO BUFFET (APERTURA Y CIERRE)
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Desayuno Buffet — Apertura y Cierre", [
        "Cómo usar esta plantilla:",
        "▸ Checklist completo para el montaje y desmontaje del buffet de desayuno.",
        "▸ Cubre desde la preparación de cocina hasta la retirada del buffet.",
        "▸ Marca ✓ cada tarea completada. Adapta a tu hotel.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables: responsable, hora, notas.",
        "▸ Ajusta las estaciones según la categoría de tu hotel.",
        "▸ Imprime en A4 y deja en cocina y en sala de desayunos.",
    ])
    create_task_sheet(wb, "Apertura Desayuno", DESAYUNO_COLOR, "Apertura del Buffet de Desayuno", [
        ("Cocina — Preparación (05:30-06:30)", [
            ("Encender hornos, planchas, baño maría y equipamiento", "Cocina"),
            ("Preparar huevos revueltos, tortilla, huevos fritos a la plancha", "Cocina"),
            ("Preparar bacon, salchichas, tomate grillado, champiñones", "Cocina"),
            ("Montar bandeja de embutidos: jamón york, pavo, chorizo, salchichón", "Cocina"),
            ("Preparar selección de quesos: manchego, brie, gouda, fresco", "Cocina"),
            ("Cortar fruta fresca de temporada: melón, sandía, piña, kiwi, fresas", "Cocina"),
            ("Preparar yogures, granola, frutos secos y miel", "Cocina"),
            ("Preparar bollería: croissants, magdalenas, pain au chocolat", "Cocina"),
            ("Preparar panes: barra, integral, centeno, tostadas, molde", "Cocina"),
            ("Preparar zumos naturales: naranja, pomelo, multifrutas", "Cocina"),
            ("Verificar temperaturas de producto caliente (>65°C) y frío (<5°C)", "Cocina"),
        ]),
        ("Sala — Montaje del buffet (06:00-06:45)", [
            ("Verificar limpieza de mesas, buffet y suelos de sala", "Sala"),
            ("Montar estación caliente: chafing dishes, fuentes, cucharas de servicio", "Buffet"),
            ("Montar estación fría: embutidos, quesos, salmón, ensaladas", "Buffet"),
            ("Montar estación de fruta y yogures con hielo decorativo", "Buffet"),
            ("Montar estación de bollería y panes con tostadora", "Buffet"),
            ("Montar estación de cereales, leche, muesli y toppings", "Buffet"),
            ("Verificar estación de bebidas: café, té, leche, zumos, agua", "Buffet"),
            ("Colocar señalización con nombre de cada plato y alérgenos", "Buffet"),
            ("Verificar mise en place de mesas: manteles, servilletas, cubiertos", "Sala"),
            ("Colocar mermeladas, mantequilla, miel en cada mesa", "Sala"),
            ("Verificar temperatura de sala (20-22°C)", "Admin"),
            ("Abrir buffet a la hora exacta según estándar del hotel", "Admin"),
        ]),
        ("Durante el servicio de desayuno", [
            ("Reponer platos calientes cada 20-30 min (no dejar vacíos)", "Buffet"),
            ("Reponer fruta, bollería y panes según consumo", "Buffet"),
            ("Mantener limpieza continua del buffet: limpiar derrames inmediatamente", "Limpieza"),
            ("Reponer zumos, café y agua en la estación de bebidas", "Buffet"),
            ("Retirar platos sucios de mesas de forma continua", "Sala"),
            ("Controlar alérgenos: atender peticiones especiales (sin gluten, vegano)", "Cocina"),
            ("Registrar conteo de comensales por tramos horarios", "Admin"),
        ]),
        ("Cierre del buffet", [
            ("Retirar señalización 10 min antes del cierre", "Buffet"),
            ("Retirar gradualmente estaciones según política del hotel", "Buffet"),
            ("Registrar sobrantes y mermas para ajustar producción futura", "Cocina"),
            ("Desmontar chafing dishes, limpiar y guardar", "Limpieza"),
            ("Limpiar y desinfectar todas las superficies del buffet", "Limpieza"),
            ("Limpiar sala de desayunos: mesas, suelos, cristales", "Limpieza"),
            ("Preparar mise en place para el día siguiente (pedidos, preparaciones)", "Cocina"),
        ]),
    ])
    create_task_sheet(wb, "Cierre Desayuno", DESAYUNO_COLOR, "Cierre Detallado del Buffet de Desayuno", [
        ("Retirada y registro de sobrantes", [
            ("Registrar sobrantes por estación: caliente, fría, bollería, fruta", "Cocina"),
            ("Separar producto reutilizable (sin contaminar) de merma", "Cocina"),
            ("Etiquetar producto reutilizable con fecha y temperatura", "Cocina"),
            ("Registrar mermas en hoja de control diario (peso y tipo)", "Cocina"),
            ("Ajustar previsión de producción para el día siguiente", "Cocina"),
            ("Fotografiar buffet antes de retirar (control de presentación)", "Admin"),
        ]),
        ("Desmontaje y limpieza de equipamiento", [
            ("Vaciar y limpiar chafing dishes (separar agua, secar)", "Limpieza"),
            ("Limpiar fuentes, bandejas y utensilios de servicio", "Limpieza"),
            ("Desinfectar superficies de la línea de buffet", "Limpieza"),
            ("Limpiar y desinfectar estación de zumos y café", "Limpieza"),
            ("Verificar que tostadoras, cafeteras y equipos están apagados", "Limpieza"),
            ("Almacenar menaje limpio en su ubicación correcta", "Limpieza"),
        ]),
        ("Sala — Limpieza y preparación", [
            ("Retirar manteles sucios y enviar a lavandería", "Sala"),
            ("Limpiar y desinfectar todas las mesas de comensales", "Limpieza"),
            ("Barrer y fregar suelo de la sala de desayunos", "Limpieza"),
            ("Limpiar cristales y espejos de la zona de buffet", "Limpieza"),
            ("Verificar sillas: limpias, alineadas y en buen estado", "Sala"),
            ("Reportar cualquier rotura o desperfecto a mantenimiento", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "01-fb-buffet-desayuno.xlsx"))


# ═══════════════════════════════════════════════════════════
# 02 — BUFFET ALMUERZO/CENA
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Buffet Almuerzo y Cena", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para montaje, reposición y desmontaje de buffet de almuerzo y cena.",
        "▸ Incluye control de calidad, temperaturas y presentación.",
        "▸ Adapta a tu tipo de hotel: resort, urbano, all-inclusive.",
        "",
        "Personalización:",
        "▸ Ajusta las estaciones según tu oferta gastronómica.",
        "▸ Para all-inclusive, duplica las estaciones que necesites.",
    ])
    create_task_sheet(wb, "Buffet Almuerzo", BUFFET_COLOR, "Buffet de Almuerzo — Montaje y Servicio", [
        ("Montaje pre-servicio", [
            ("Verificar limpieza de línea de buffet y mesas de apoyo", "Limpieza"),
            ("Montar estación de ensaladas y fríos: lechuga, tomate, quinoa, pasta fría", "Buffet"),
            ("Montar estación de entrantes fríos: gazpacho, carpaccio, salmón", "Buffet"),
            ("Montar estación caliente: arroces, pasta, carnes, pescados", "Buffet"),
            ("Montar estación de guarniciones: patatas, verduras, legumbres", "Buffet"),
            ("Montar show cooking / estación de acción (plancha, wok, pasta)", "Buffet"),
            ("Montar estación de postres: tartas, fruta, helados, lácteos", "Buffet"),
            ("Verificar señalización con nombre de plato y alérgenos (14 obligatorios)", "Buffet"),
            ("Colocar cubiertos de servicio limpios en cada fuente", "Buffet"),
            ("Verificar temperaturas: caliente >65°C, frío <5°C (registrar APPCC)", "Buffet"),
        ]),
        ("Reposición durante servicio", [
            ("Ronda de reposición cada 15-20 min (nunca dejar fuentes vacías)", "Buffet"),
            ("Cambiar fuentes completas (no rellenar sobre comida existente)", "Buffet"),
            ("Mantener show cooking activo con chef visible y producto fresco", "Buffet"),
            ("Reponer bebidas: agua, zumos, refrescos, vino de la casa", "Buffet"),
            ("Limpiar derrames inmediatamente y mantener línea presentable", "Limpieza"),
            ("Registrar reposiciones para ajustar producción (qué se acaba primero)", "Cocina"),
        ]),
        ("Cierre del buffet", [
            ("Retirar estaciones gradualmente (señalización primero)", "Buffet"),
            ("Registrar sobrantes por estación para control de mermas", "Cocina"),
            ("Desmontar chafing dishes, limpiar y almacenar", "Limpieza"),
            ("Limpiar línea de buffet, suelos y mesas de apoyo", "Limpieza"),
        ]),
    ])
    create_task_sheet(wb, "Buffet Cena", BUFFET_COLOR, "Buffet de Cena — Montaje y Servicio", [
        ("Montaje y diferencias vs. almuerzo", [
            ("Ajustar iluminación de sala a ambiente de cena (más cálida)", "Sala"),
            ("Montar estación de entrantes premium: foie, tartar, ceviche", "Buffet"),
            ("Montar estación caliente: platos de cena (más elaborados que almuerzo)", "Buffet"),
            ("Montar show cooking especial de cena: parrilla, robata, flambé", "Buffet"),
            ("Montar estación de quesos con selección de vinos", "Buffet"),
            ("Montar estación de postres elaborados y petit fours", "Buffet"),
            ("Verificar carta de vinos disponible en sala (servicio a mesa)", "Sala"),
            ("Colocar velas o iluminación decorativa según estándar", "Sala"),
        ]),
        ("Servicio y cierre", [
            ("Reposición igual que almuerzo: cada 15-20 min, fuentes completas", "Buffet"),
            ("Coordinar con bar del hotel para servicio de cócteles pre-cena", "Sala"),
            ("Ofrecer café, infusiones y digestivos al finalizar", "Sala"),
            ("Cierre y desmontaje siguiendo mismo protocolo que almuerzo", "Buffet"),
            ("Preparar mise en place del desayuno siguiente si hay turno cruzado", "Cocina"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "02-fb-buffet-comida-cena.xlsx"))


# ═══════════════════════════════════════════════════════════
# 03 — RESTAURANTE À LA CARTE
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Restaurante À la Carte — Apertura y Cierre", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para apertura y cierre del restaurante à la carte del hotel.",
        "▸ Cubre reservas, montaje de sala, servicio y cierre.",
        "▸ Ideal para restaurante gastronómico, grill o temático.",
        "",
        "Personalización:",
        "▸ Ajusta las tareas a tu estilo de restaurante.",
        "▸ Imprime y deja en el office de sala.",
    ])
    create_task_sheet(wb, "Apertura À la Carte", BANQUETE_COLOR,
                      "Apertura del Restaurante À la Carte", [
        ("Preparación de sala (16:00-18:00)", [
            ("Revisar reservas del día: número, horario, alergias, VIPs", "À la Carte"),
            ("Consultar lista de alérgenos y peticiones especiales (celiaco, vegano)", "À la Carte"),
            ("Montar mesas según plano de reservas y distribución de sala", "À la Carte"),
            ("Verificar stock de vinos por copa y actualizar carta si agotados", "À la Carte"),
            ("Preparar mise en place de camareros: cubiertos, cristalería, servilletas", "À la Carte"),
            ("Pulir cristalería y cubertería de todas las mesas", "À la Carte"),
            ("Ajustar iluminación y música ambiental según estándar", "À la Carte"),
            ("Briefing de equipo: carta del día, VIPs, sugerencias, vinos, protocolo", "À la Carte"),
        ]),
        ("Control pre-apertura", [
            ("Verificar funcionamiento de TPV y sistema de cargo a habitación", "Admin"),
            ("Comprobar temperatura de sala (20-22°C) y ventilación", "Admin"),
            ("Revisar limpieza de baños de clientes", "Limpieza"),
            ("Confirmar con cocina: mise en place lista, platos agotados, tiempos", "Cocina"),
            ("Verificar estación de pan, mantequilla y aceite de oliva", "À la Carte"),
            ("Abrir restaurante puntualmente según horario publicado", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Cierre À la Carte", BANQUETE_COLOR,
                      "Cierre del Restaurante À la Carte", [
        ("Cierre de servicio", [
            ("Cuadrar caja y TPV: efectivo, tarjeta, cargos a habitación", "Admin"),
            ("Repasar mesas: recoger mantelería, cristalería, cubertería", "À la Carte"),
            ("Reponer office de camareros: servilletas, cubiertos, condimentos", "À la Carte"),
            ("Limpiar barra de bar y recoger cristalería", "Limpieza"),
            ("Apagar cafeteras, vitrinas de vino y equipos de sala", "À la Carte"),
            ("Cerrar botellas abiertas, registrar mermas de vino/bebida", "À la Carte"),
            ("Feedback con cocina: platos más vendidos, incidencias, tiempos", "Cocina"),
        ]),
        ("Limpieza y preparación siguiente día", [
            ("Limpiar y desinfectar todas las mesas y sillas", "Limpieza"),
            ("Barrer y fregar suelos de sala y entrada", "Limpieza"),
            ("Verificar baños de clientes: limpieza final", "Limpieza"),
            ("Revisar reservas del día siguiente y preparar plano de sala", "Admin"),
            ("Apagar luces, música y climatización al salir", "Admin"),
            ("Cerrar puertas y activar alarma según protocolo", "Seguridad"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "03-fb-restaurante-carte.xlsx"))


# ═══════════════════════════════════════════════════════════
# 04 — F&B OUTLETS (Pool Bar, Lobby Bar, Snack Bar)
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · F&B Outlets — Pool Bar, Lobby Bar, Snack Bar", [
        "Cómo usar esta plantilla:",
        "▸ Checklists para los 3 outlets satélite más comunes en hoteles.",
        "▸ Pool Bar, Lobby Bar / Lounge y Snack Bar / Grab & Go.",
        "▸ Adapta según los outlets de tu hotel.",
        "",
        "Personalización:",
        "▸ Activa solo las pestañas de outlets que tenga tu hotel.",
        "▸ Ajusta horarios y oferta a tu categoría (4*, 5*, resort).",
    ])
    create_task_sheet(wb, "Pool Bar", PISCINA_COLOR,
                      "Pool Bar — Apertura, Servicio y Cierre", [
        ("Apertura Pool Bar", [
            ("Montar barra: limpiar superficies, colocar equipos, hielo", "Pool Bar"),
            ("Verificar stock de cervezas, refrescos y licores base para cócteles", "Pool Bar"),
            ("Preparar hielo suficiente para el turno (cubitera grande + repuesto)", "Pool Bar"),
            ("Colocar cartas/pizarras de bebidas y snacks con precios visibles", "Pool Bar"),
            ("Verificar TPV portátil operativo y sistema de cargo a habitación", "Pool Bar"),
            ("Comprobar aforo y sistema de pulseras/toallas AI si aplica", "Pool Bar"),
        ]),
        ("Servicio activo", [
            ("Atender pedidos en barra y servicio a tumbonas", "Pool Bar"),
            ("Mantener hielo, reponer bebidas frías según consumo", "Pool Bar"),
            ("Preparar cócteles y bebidas especiales según carta", "Pool Bar"),
            ("Recoger vasos y platos vacíos de zona de tumbonas", "Pool Bar"),
            ("Controlar consumo de huéspedes (pulsera/cargo habitación)", "Pool Bar"),
            ("Mantener barra y zona de servicio limpia en todo momento", "Limpieza"),
        ]),
        ("Cierre Pool Bar", [
            ("Cuadrar caja y registrar cargos a habitación pendientes", "Admin"),
            ("Recoger cristalería y menaje de toda la zona de piscina", "Pool Bar"),
            ("Limpiar barra, equipos y superficies a fondo", "Limpieza"),
            ("Almacenar licores y producto perecedero en cámara", "Pool Bar"),
            ("Cubrir equipamiento y cerrar acceso según protocolo", "Pool Bar"),
        ]),
    ])
    create_task_sheet(wb, "Lobby Bar - Lounge", EVENTO_COLOR,
                      "Lobby Bar / Lounge — Servicio Diario", [
        ("Apertura y servicio continuo", [
            ("Preparar estación de café y té continuo (24h o según horario)", "Lobby Bar"),
            ("Montar oferta de snacks premium: frutos secos, aceitunas, chips", "Lobby Bar"),
            ("Verificar carta de cócteles actualizada y stock de ingredientes", "Lobby Bar"),
            ("Configurar playlist/música ambiental según franja horaria", "Lobby Bar"),
            ("Mantener zona impecable: limpieza continua de mesas y barra", "Limpieza"),
            ("Ofrecer servicio discreto y atento a huéspedes del lobby", "Lobby Bar"),
        ]),
        ("Cierre Lobby Bar", [
            ("Cuadrar caja del turno y revisar cargos a habitación", "Admin"),
            ("Recoger cristalería y limpiar todas las mesas", "Limpieza"),
            ("Limpiar máquina de café y equipos de barra", "Limpieza"),
            ("Almacenar perecederos y cerrar botellas abiertas", "Lobby Bar"),
            ("Dejar zona presentable para el turno de noche/mañana", "Lobby Bar"),
        ]),
    ])
    create_task_sheet(wb, "Snack Bar - Grab Go", BUFFET_COLOR,
                      "Snack Bar / Grab & Go — Operativa Diaria", [
        ("Montaje y reposición", [
            ("Reponer vitrina de sándwiches, ensaladas y wraps frescos", "Snack Bar"),
            ("Verificar temperatura de vitrina refrigerada (<5°C)", "Snack Bar"),
            ("Reponer stock de bebidas: agua, zumos, refrescos, smoothies", "Snack Bar"),
            ("Verificar estación de autoservicio de café operativa", "Snack Bar"),
            ("Aplicar FIFO: rotar productos y retirar caducados", "Snack Bar"),
            ("Actualizar precios y cartelería si hay cambios", "Snack Bar"),
        ]),
        ("Control y cierre", [
            ("Registrar mermas y productos retirados por caducidad", "Admin"),
            ("Limpiar vitrinas, superficies y zona de autoservicio", "Limpieza"),
            ("Cuadrar caja y registrar ventas del día", "Admin"),
            ("Preparar pedido de producto fresco para el día siguiente", "Snack Bar"),
            ("Verificar temperaturas de cierre y registrar en hoja APPCC", "Snack Bar"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "04-fb-outlets.xlsx"))


# ═══════════════════════════════════════════════════════════
# 05 — ROOM SERVICE Y MINIBAR
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Room Service y Minibar", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para la operativa de room service y minibar.",
        "▸ Room service: preparación, servicio, recogida.",
        "▸ Minibar: reposición, control, facturación.",
        "",
        "Personalización:",
        "▸ Ajusta el menú y los horarios a tu operación.",
        "▸ Añade protocolos VIP si tu hotel los requiere.",
    ])
    create_task_sheet(wb, "Room Service", ROOMSERVICE_COLOR,
                      "Room Service — Operativa Completa", [
        ("Preparación del turno", [
            ("Verificar carta de room service actualizada (impresa y digital)", "Room Service"),
            ("Comprobar stock de minibar reposición: agua, refrescos, snacks", "Room Service"),
            ("Verificar bandejas, cloches, cubiertos de plata, servilletas", "Room Service"),
            ("Revisar amenities: flores, tarjeta del chef, chocolate de bienvenida", "Room Service"),
            ("Verificar carrito de servicio: mantel, vela, cubiertos, condimentos", "Room Service"),
            ("Comprobar teléfono de pedidos y sistema de PMS/POS operativo", "Room Service"),
            ("Revisar pedidos pendientes y pre-pedidos del día (despertadores, etc.)", "Room Service"),
        ]),
        ("Recepción y preparación de pedidos", [
            ("Anotar pedido completo: habitación, nº comensales, alérgenos, hora", "Room Service"),
            ("Confirmar tiempo estimado de entrega al huésped (máx. 30-45 min)", "Room Service"),
            ("Pasar comanda a cocina con prioridad y notas especiales", "Cocina"),
            ("Preparar bandeja/carrito según estándar: mantel, cubertería, condimentos", "Room Service"),
            ("Verificar presentación del plato antes de salir de cocina", "Room Service"),
            ("Controlar temperatura: plato caliente con cloche, frío con hielo", "Room Service"),
            ("Incluir sal, pimienta, salsas según el plato", "Room Service"),
        ]),
        ("Servicio en habitación", [
            ("Llamar a la puerta según protocolo: 3 golpes + 'Room service'", "Room Service"),
            ("Presentar el pedido y ofrecer montaje en mesa o escritorio", "Room Service"),
            ("Abrir botellas de vino/cava si aplica (ofrecer cata al huésped)", "Room Service"),
            ("Confirmar que todo es correcto y ofrecer cualquier cosa adicional", "Room Service"),
            ("Dejar tarjeta con número de recogida en la bandeja", "Room Service"),
            ("Registrar hora de entrega y habitación en el sistema", "Room Service"),
        ]),
        ("Recogida y cierre", [
            ("Ronda de recogida de bandejas cada 60-90 min por pasillos", "Room Service"),
            ("Verificar bandejas recogidas: conteo de cubertería y cristalería", "Room Service"),
            ("Limpiar y desinfectar carritos y bandejas", "Limpieza"),
            ("Inventariar cubertería de plata al final del turno", "Room Service"),
            ("Registrar incidencias, quejas o peticiones especiales", "Admin"),
            ("Preparar reporte de room service del turno: pedidos, tiempos, incidencias", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Minibar", MINIBAR_COLOR,
                      "Minibar — Reposición y Control", [
        ("Ronda de reposición", [
            ("Obtener listado de habitaciones check-out y check-in del día", "Minibar"),
            ("Preparar carrito de reposición con par stock completo", "Minibar"),
            ("Reponer habitaciones check-out: inventario completo + limpieza", "Minibar"),
            ("Reponer habitaciones ocupadas según consumo registrado", "Minibar"),
            ("Verificar caducidades de productos del minibar", "Minibar"),
            ("Registrar consumos en sistema PMS para facturación", "Minibar"),
            ("Colocar carta de minibar actualizada con precios", "Minibar"),
        ]),
        ("Control y reporting", [
            ("Inventario semanal de stock de minibar en almacén", "Minibar"),
            ("Verificar temperatura de minibar en habitaciones (spot check)", "Minibar"),
            ("Reportar averías de minibares a mantenimiento", "Admin"),
            ("Analizar consumo por tipo de producto (tendencias, rotación)", "Admin"),
            ("Gestionar pedidos de reposición a almacén central", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "05-fb-room-service-minibar.xlsx"))


# ═══════════════════════════════════════════════════════════
# 06 — BANQUETES Y EVENTOS
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Banquetes y Eventos del Hotel", [
        "Cómo usar esta plantilla:",
        "▸ Checklists para los principales tipos de evento en hotel.",
        "▸ Bodas, convenciones, cenas de gala, coffee breaks.",
        "▸ Usa la pestaña del evento que tengas más próximo.",
        "",
        "Personalización:",
        "▸ Combina con los checklists generales (01-05) para cobertura completa.",
        "▸ Adapta según la categoría y los salones de tu hotel.",
    ])
    create_task_sheet(wb, "Banquetes y Eventos", EVENTO_COLOR,
                      "Banquetes y Eventos — Checklist Integral", [
        ("Coordinación pre-evento", [
            ("Confirmar BEO (Banquet Event Order) con cliente/wedding planner", "Banquetes"),
            ("Verificar montaje de salón según plano aprobado (teatro, imperial, cóctel)", "Banquetes"),
            ("Coordinar con departamentos: recepción, housekeeping, AV, seguridad", "Admin"),
            ("Confirmar menú definitivo, alérgenos y distribución de mesas", "Cocina"),
            ("Verificar equipamiento AV: proyector, micro, pantalla, traducción", "Admin"),
            ("Coordinar timing con DJ, fotógrafo, florista u otros proveedores", "Admin"),
            ("Preparar habitaciones especiales con amenities si aplica (novios, VIP)", "Admin"),
        ]),
        ("Montaje de boda / cena de gala", [
            ("Montar mesa imperial o mesas redondas según BEO con mantelería", "Banquetes"),
            ("Colocar cubertería completa para todos los tiempos del menú", "Banquetes"),
            ("Colocar centro de mesa, tarjetas de asiento y menú impreso", "Banquetes"),
            ("Preparar cóctel de bienvenida: canapés, cava, zona de terraza/jardín", "Banquetes"),
            ("Servicio de mesa sincronizado: todos los platos al mismo tiempo", "Banquetes"),
            ("Coordinar corte de tarta/discursos con fotógrafo y DJ", "Banquetes"),
            ("Barra libre según horario pactado (controlar cierre)", "Banquetes"),
            ("Servicio de recena/snack nocturno si aplica", "Banquetes"),
        ]),
        ("Convención / Congreso", [
            ("Verificar montaje de sala plenaria según formato BEO", "Banquetes"),
            ("Montar coffee break mañana: café, zumos, bollería, fruta, mini-sándwiches", "Banquetes"),
            ("Servir almuerzo según formato: buffet, box lunch o sentado", "Banquetes"),
            ("Coffee break tarde (variar vs. mañana): snacks, smoothies, fruta", "Banquetes"),
            ("Mantener estación de agua y café disponible todo el día", "Banquetes"),
            ("Cocktail networking final: canapés, barra de cócteles", "Banquetes"),
        ]),
        ("Desmontaje y cierre", [
            ("Desmontar salón según protocolo: mesas, sillas, decoración, AV", "Banquetes"),
            ("Inventariar cubertería, cristalería y menaje (reportar roturas)", "Banquetes"),
            ("Limpiar salón a fondo: suelo, moqueta, paredes, baños", "Limpieza"),
            ("Cerrar caja de banquetes y conciliar con factura del evento", "Admin"),
            ("Registrar feedback del cliente y enviar encuesta de satisfacción", "Admin"),
            ("Late check-out coordinado con recepción si procede", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "06-fb-banquetes-eventos.xlsx"))


# ═══════════════════════════════════════════════════════════
# 07 — RECEPCIÓN Y TURNOS
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Recepción — Turnos y Protocolo Check-in/out", [
        "Cómo usar esta plantilla:",
        "▸ 4 checklists para los 3 turnos de recepción + protocolo check-in/out.",
        "▸ Cubre desde la auditoría nocturna hasta la gestión de VIPs.",
        "▸ Imprime el turno correspondiente y deja en el mostrador.",
        "",
        "Personalización:",
        "▸ Ajusta según tu PMS (Opera, Mews, Protel, etc.).",
        "▸ Añade los procedimientos específicos de tu cadena/marca.",
    ])
    create_task_sheet(wb, "Turno Mañana", RECEPCION_COLOR,
                      "Recepción — Turno Mañana (07:00-15:00)", [
        ("Inicio de turno", [
            ("Leer logbook del turno nocturno: incidencias, peticiones, notas", "Recepción"),
            ("Revisar listado de check-outs del día y facturas pendientes", "Recepción"),
            ("Revisar listado de llegadas y pre-asignación de habitaciones", "Recepción"),
            ("Consultar estado de habitaciones en PMS con housekeeping", "Recepción"),
            ("Verificar caja del turno: fondo, cambio, datáfono operativo", "Recepción"),
        ]),
        ("Operativa del turno", [
            ("Gestionar check-outs: factura, minibar, feedback, transporte", "Recepción"),
            ("Atender consultas de huéspedes: información, reservas, reclamaciones", "Recepción"),
            ("Coordinar con bell boys para equipajes de salidas y llegadas", "Recepción"),
            ("Actualizar rack de habitaciones según salidas confirmadas", "Recepción"),
            ("Coordinar con housekeeping las prioridades de limpieza", "Recepción"),
            ("Gestionar early check-ins si hay habitaciones disponibles", "Recepción"),
            ("Registrar todas las incidencias en el logbook para turno tarde", "Recepción"),
        ]),
    ])
    create_task_sheet(wb, "Turno Tarde", RECEPCION_COLOR,
                      "Recepción — Turno Tarde (15:00-23:00)", [
        ("Inicio de turno", [
            ("Leer logbook del turno mañana: check-outs, incidencias, notas", "Recepción"),
            ("Revisar llegadas pendientes y habitaciones asignadas", "Recepción"),
            ("Verificar VIPs del día y amenities de bienvenida preparados", "Recepción"),
            ("Verificar caja del turno y estado del datáfono", "Recepción"),
        ]),
        ("Operativa del turno", [
            ("Gestionar check-ins: registro, llaves, información de servicios/wifi/parking", "Recepción"),
            ("Ofrecer upselling de categoría de habitación cuando proceda", "Recepción"),
            ("Gestionar walk-ins y overbooking según protocolo", "Recepción"),
            ("Atender reclamaciones y gestionar cambios de habitación", "Recepción"),
            ("Gestionar late check-outs según disponibilidad y política", "Recepción"),
            ("Confirmar reservas del día siguiente y preparar documentación", "Recepción"),
            ("Preparar documentación para auditoría nocturna", "Recepción"),
            ("Registrar incidencias en logbook para turno noche", "Recepción"),
        ]),
    ])
    create_task_sheet(wb, "Turno Noche", RECEPCION_COLOR,
                      "Recepción — Turno Noche / Auditoría (23:00-07:00)", [
        ("Night Audit", [
            ("Ejecutar night audit en PMS: cierre de día contable", "Recepción"),
            ("Verificar cargos de restaurantes, spa, minibar cargados correctamente", "Recepción"),
            ("Generar informes: ocupación, ADR, RevPAR, producción por departamento", "Recepción"),
            ("Procesar no-shows: aplicar penalizaciones según política", "Recepción"),
            ("Preparar listado de salidas y llegadas del día siguiente", "Recepción"),
        ]),
        ("Operativa nocturna", [
            ("Atender recepción nocturna: late arrivals, peticiones, emergencias", "Recepción"),
            ("Realizar ronda de seguridad visual del lobby y zonas comunes", "Seguridad"),
            ("Verificar puertas de acceso y cámaras de seguridad", "Seguridad"),
            ("Preparar logbook completo para turno de mañana", "Recepción"),
            ("Verificar pre-asignación de habitaciones para llegadas del día", "Recepción"),
            ("Coordinar early breakfast / madrugadores con F&B si aplica", "Recepción"),
        ]),
    ])
    create_task_sheet(wb, "Check-in Check-out", RECEPCION_COLOR,
                      "Protocolo de Check-in y Check-out", [
        ("Protocolo de Check-in", [
            ("Saludar al huésped por su nombre (consultar reserva)", "Recepción"),
            ("Solicitar documentación de identidad y tarjeta de crédito", "Recepción"),
            ("Confirmar detalles de la reserva: fechas, tipo de habitación, régimen", "Recepción"),
            ("Informar de servicios del hotel: restaurante, spa, wifi, parking, horarios", "Recepción"),
            ("Ofrecer upgrade/upselling de habitación si hay disponibilidad", "Recepción"),
            ("Entregar llaves, mapa del hotel y tarjeta de bienvenida", "Recepción"),
            ("Ofrecer acompañamiento a la habitación por bell boy", "Recepción"),
        ]),
        ("Protocolo de Check-out", [
            ("Preguntar si ha consumido minibar y verificar si es necesario", "Recepción"),
            ("Presentar factura desglosada y verificar todos los cargos", "Recepción"),
            ("Preguntar por la estancia y solicitar feedback", "Recepción"),
            ("Ofrecer ayuda con equipaje y transporte al aeropuerto/estación", "Recepción"),
            ("Invitar a dejar review online (TripAdvisor, Google, Booking)", "Recepción"),
            ("Despedir cordialmente e invitar a volver", "Recepción"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "07-recepcion-turnos.xlsx"))


# ═══════════════════════════════════════════════════════════
# 08 — GUEST SERVICES
# ═══════════════════════════════════════════════════════════
def gen_08():
    wb = Workbook()
    add_instructions_sheet(wb, "08 · Guest Services — Conserjería y Guest Experience", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para conserjería y gestión de experiencia del huésped.",
        "▸ Reservas externas, transfers, peticiones especiales, fidelización.",
        "▸ Ideal para hoteles 4-5* con servicio personalizado.",
        "",
        "Personalización:",
        "▸ Ajusta los servicios a la oferta de tu destino.",
        "▸ Añade proveedores locales de confianza.",
    ])
    create_task_sheet(wb, "Conserjería", RECEPCION_COLOR,
                      "Conserjería — Servicios al Huésped", [
        ("Gestión de peticiones", [
            ("Gestionar reservas en restaurantes externos (confirmar y enviar detalles)", "Conserjería"),
            ("Coordinar transfers aeropuerto/estación con empresa de transporte", "Conserjería"),
            ("Organizar excursiones, tours y actividades para huéspedes", "Conserjería"),
            ("Gestionar peticiones especiales: flores, champán, tartas de cumpleaños", "Conserjería"),
            ("Mantener directorio actualizado de proveedores locales de confianza", "Conserjería"),
            ("Coordinar servicio de equipajes con bell boys (llegadas y salidas)", "Conserjería"),
            ("Gestionar servicio de lavandería y tintorería de huéspedes", "Conserjería"),
        ]),
        ("Comunicación y seguimiento", [
            ("Confirmar todas las reservas y servicios gestionados al huésped", "Conserjería"),
            ("Seguimiento de peticiones en curso: estado y resolución", "Conserjería"),
            ("Registrar todas las gestiones en el PMS/CRM del huésped", "Admin"),
            ("Preparar información personalizada: mapas, horarios, recomendaciones", "Conserjería"),
            ("Coordinar con recepción para amenities de bienvenida VIP", "Conserjería"),
        ]),
    ])
    create_task_sheet(wb, "Guest Experience", RECEPCION_COLOR,
                      "Guest Experience / Fidelización", [
        ("Gestión proactiva", [
            ("Revisar listado de huéspedes repetidores y VIPs del día", "Guest Exp."),
            ("Preparar amenities de bienvenida personalizados según perfil CRM", "Guest Exp."),
            ("Follow-up de incidencias abiertas: verificar resolución y satisfacción", "Guest Exp."),
            ("Monitorizar reviews online: TripAdvisor, Google, Booking (responder)", "Guest Exp."),
            ("Enviar encuesta post-estancia a huéspedes del día anterior", "Guest Exp."),
            ("Actualizar preferencias en CRM: habitación, almohada, minibar, dieta", "Guest Exp."),
        ]),
        ("Programas de fidelización", [
            ("Identificar huéspedes elegibles para programa de fidelidad", "Guest Exp."),
            ("Preparar detalles especiales: cumpleaños, aniversarios, luna de miel", "Guest Exp."),
            ("Coordinar con F&B amenities gastronómicas: chocolate, fruta, vino", "Guest Exp."),
            ("Generar informe mensual de satisfacción: NPS, puntuación media", "Admin"),
            ("Proponer mejoras basadas en feedback recurrente de huéspedes", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "08-guest-services.xlsx"))


# ═══════════════════════════════════════════════════════════
# 09 — HOUSEKEEPING
# ═══════════════════════════════════════════════════════════
def gen_09():
    wb = Workbook()
    add_instructions_sheet(wb, "09 · Housekeeping — 5 Protocolos de Limpieza", [
        "Cómo usar esta plantilla:",
        "▸ 5 checklists: Checkout, Estancia, Turndown, Deep Cleaning, Lencería.",
        "▸ Protocolo paso a paso para cada tipo de limpieza de habitación.",
        "▸ Imprime y entrega a cada camarera de pisos.",
        "",
        "Personalización:",
        "▸ Ajusta a la categoría de tu hotel (3*, 4*, 5*).",
        "▸ Añade estándares de marca/cadena si aplica.",
    ])
    create_task_sheet(wb, "Limpieza Checkout", HOUSEKEEPING_COLOR,
                      "Limpieza de Habitación — Checkout", [
        ("Preparación y retirada", [
            ("Llamar/tocar antes de entrar según protocolo", "Housekeeping"),
            ("Retirar toda la ropa de cama: sábanas, fundas, almohadas", "Housekeeping"),
            ("Revisar si hay objetos olvidados (debajo cama, armario, caja fuerte, baño)", "Housekeeping"),
            ("Retirar basura de todas las papeleras y bolsas", "Housekeeping"),
            ("Retirar amenities usados del baño y toallas sucias", "Housekeeping"),
        ]),
        ("Limpieza completa", [
            ("Limpiar y desinfectar baño completo: sanitarios, ducha, grifería, espejo", "Housekeeping"),
            ("Limpiar mobiliario: escritorio, mesillas, armario interior, minibar exterior", "Housekeeping"),
            ("Colocar amenities nuevos: gel, champú, jabón, gorro, kit dental", "Housekeeping"),
            ("Hacer cama completa: sábanas, funda nórdica, almohadas, colcha", "Housekeeping"),
            ("Aspirar/fregar suelo completo incluyendo debajo de muebles", "Housekeeping"),
        ]),
        ("Verificación final", [
            ("Verificar funcionamiento: luces, AC/calefacción, TV, teléfono", "Housekeeping"),
            ("Verificar minibar: repuesto completo según par stock", "Housekeeping"),
            ("Verificar perchas, albornoz, zapatillas, carta de servicios", "Housekeeping"),
            ("Comprobar olor y ventilación de la habitación", "Housekeeping"),
            ("Reportar habitación como limpia en PMS/app", "Housekeeping"),
            ("Reportar cualquier desperfecto o avería a mantenimiento", "Housekeeping"),
        ]),
    ])
    create_task_sheet(wb, "Limpieza Estancia", HOUSEKEEPING_COLOR,
                      "Limpieza de Habitación — Estancia (Ocupada)", [
        ("Protocolo de entrada", [
            ("Llamar a la puerta y esperar respuesta antes de entrar", "Housekeeping"),
            ("Respetar cartel 'No molestar' y registrar para seguimiento", "Housekeeping"),
            ("No mover objetos personales del huésped", "Housekeeping"),
        ]),
        ("Limpieza de estancia", [
            ("Hacer cama sin cambiar sábanas (excepto programa green/petición)", "Housekeeping"),
            ("Vaciar papeleras y retirar basura", "Housekeeping"),
            ("Limpiar baño: reponer toallas usadas, amenities, limpiar sanitarios", "Housekeeping"),
            ("Aspirar suelo y ordenar sin mover objetos personales", "Housekeeping"),
            ("Reponer amenities gastados: café, té, azúcar, agua de cortesía", "Housekeeping"),
            ("Reportar cualquier daño o necesidad de mantenimiento", "Housekeeping"),
        ]),
    ])
    create_task_sheet(wb, "Turndown Service", HOUSEKEEPING_COLOR,
                      "Turndown Service — Servicio de Cobertura", [
        ("Preparación de habitación para la noche", [
            ("Retirar colcha y cojines decorativos, doblar y guardar", "Housekeeping"),
            ("Abrir cama en ángulo (45°) y colocar zapatillas/albornoz", "Housekeeping"),
            ("Colocar chocolate/bombón en la almohada", "Housekeeping"),
            ("Cerrar cortinas y persianas", "Housekeeping"),
            ("Encender luz de mesilla y apagar luces generales", "Housekeeping"),
            ("Vaciar papeleras si están llenas", "Housekeeping"),
            ("Reponer toallas de baño usadas y colocar alfombrilla", "Housekeeping"),
            ("Ajustar AC a temperatura nocturna según estándar (20-22°C)", "Housekeeping"),
        ]),
    ])
    create_task_sheet(wb, "Deep Cleaning", HOUSEKEEPING_COLOR,
                      "Deep Cleaning — Limpieza Profunda Periódica", [
        ("Limpieza profunda (trimestral o según rotación)", [
            ("Mover todo el mobiliario y limpiar detrás y debajo", "Housekeeping"),
            ("Limpiar rejillas de AC, extractores y conductos accesibles", "Housekeeping"),
            ("Desmontar y limpiar cortinas (enviar a lavandería)", "Housekeeping"),
            ("Desinfectar colchón con vapor y verificar estado", "Housekeeping"),
            ("Limpiar cristales interiores y exteriores de ventanas", "Housekeeping"),
            ("Descalcificar grifería, ducha y desagües del baño", "Housekeeping"),
            ("Limpiar minibar por dentro: descongelar, desinfectar, secar", "Housekeeping"),
            ("Pulir suelos (parqué, mármol, moqueta según material)", "Housekeeping"),
        ]),
    ])
    create_task_sheet(wb, "Control Lencería", HOUSEKEEPING_COLOR,
                      "Control de Lencería y Lavandería", [
        ("Gestión diaria de lencería", [
            ("Contar ropa sucia por tipo: sábanas, fundas, toallas, albornoces", "Lencería"),
            ("Separar piezas manchadas o dañadas para tratamiento especial", "Lencería"),
            ("Enviar a lavandería industrial con albarán de conteo", "Lencería"),
            ("Recepcionar lencería limpia y verificar conteo vs. albarán", "Lencería"),
            ("Almacenar por tipo y tamaño manteniendo par stock por planta", "Lencería"),
            ("Reportar piezas deterioradas para reposición y baja de inventario", "Lencería"),
        ]),
        ("Control semanal", [
            ("Inventario semanal completo por tipo de pieza", "Lencería"),
            ("Verificar par stock mínimo (3 juegos por habitación)", "Lencería"),
            ("Revisar calidad de lavado: manchas, olor, suavidad", "Lencería"),
            ("Actualizar registro de bajas y necesidades de compra", "Admin"),
            ("Coordinar con compras pedido de reposición si necesario", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "09-housekeeping.xlsx"))


# ═══════════════════════════════════════════════════════════
# 10 — ÁREAS PÚBLICAS
# ═══════════════════════════════════════════════════════════
def gen_10():
    wb = Workbook()
    add_instructions_sheet(wb, "10 · Áreas Públicas — Lobby, Pasillos, Baños, Parking", [
        "Cómo usar esta plantilla:",
        "▸ 4 checklists para la limpieza y mantenimiento de zonas comunes.",
        "▸ Lobby, pasillos/ascensores, baños públicos y parking.",
        "▸ Ideal para supervisores de pisos y áreas públicas.",
        "",
        "Personalización:",
        "▸ Ajusta la frecuencia según la categoría y ocupación de tu hotel.",
        "▸ Añade zonas específicas de tu hotel (jardín, terraza, piscina).",
    ])
    create_task_sheet(wb, "Lobby", AREAS_PUBLICAS_COLOR,
                      "Lobby — Limpieza y Mantenimiento", [
        ("Limpieza diaria (mínimo 2 veces/día)", [
            ("Aspirar/fregar suelo del lobby completo (mármol, moqueta)", "Lobby"),
            ("Limpiar cristales de puertas de entrada y escaparates", "Lobby"),
            ("Limpiar mostradores de recepción y conserjería", "Lobby"),
            ("Limpiar y ordenar mobiliario de zona de espera", "Lobby"),
            ("Verificar flores frescas y reemplazar si están marchitas", "Lobby"),
            ("Verificar ambientador/difusor de aroma funcionando", "Lobby"),
            ("Vaciar papeleras y ceniceros exteriores de la entrada", "Lobby"),
        ]),
    ])
    create_task_sheet(wb, "Pasillos y Ascensores", AREAS_PUBLICAS_COLOR,
                      "Pasillos y Ascensores — Todas las Plantas", [
        ("Limpieza diaria", [
            ("Aspirar moqueta de pasillos en todas las plantas", "Pasillos"),
            ("Limpiar marcos de puertas, zócalos y manillas de habitaciones", "Pasillos"),
            ("Verificar señalética y numeración de habitaciones (visible y limpia)", "Pasillos"),
            ("Limpiar espejos, puertas y botonera de ascensores", "Pasillos"),
            ("Verificar iluminación: reemplazar bombillas fundidas", "Pasillos"),
            ("Retirar bandejas de room service abandonadas en pasillos", "Pasillos"),
            ("Reportar daños en paredes, moquetas o mobiliario a mantenimiento", "Pasillos"),
        ]),
    ])
    create_task_sheet(wb, "Baños Públicos", AREAS_PUBLICAS_COLOR,
                      "Baños Públicos — Limpieza y Reposición", [
        ("Limpieza periódica (cada 2-3 horas)", [
            ("Desinfectar sanitarios: inodoros, urinarios, lavabos", "Baños"),
            ("Reponer jabón, papel higiénico y ambientador", "Baños"),
            ("Limpiar espejos y grifería hasta brillo", "Baños"),
            ("Fregar suelo con producto desinfectante", "Baños"),
            ("Vaciar papeleras y contenedores de higiene femenina", "Baños"),
            ("Verificar secadores de manos y dispensadores funcionan", "Baños"),
            ("Firmar hoja de control horario de limpieza (visible)", "Baños"),
        ]),
    ])
    create_task_sheet(wb, "Parking", AREAS_PUBLICAS_COLOR,
                      "Parking — Mantenimiento y Seguridad", [
        ("Control diario", [
            ("Barrer/soplar hojas y basura de las plazas de aparcamiento", "Parking"),
            ("Verificar iluminación de todas las zonas del parking", "Parking"),
            ("Comprobar señalización horizontal (líneas) y vertical (carteles)", "Parking"),
            ("Verificar barrera/acceso automático y lectores de tarjeta", "Parking"),
            ("Verificar extintores y señalización de evacuación", "Parking"),
            ("Limpiar manchas de aceite o líquidos del suelo", "Parking"),
            ("Identificar y reportar vehículos abandonados o sin autorización", "Parking"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "10-areas-publicas.xlsx"))


# ═══════════════════════════════════════════════════════════
# 11 — PISCINA
# ═══════════════════════════════════════════════════════════
def gen_11():
    wb = Workbook()
    add_instructions_sheet(wb, "11 · Piscina — Apertura, Cierre, Mantenimiento y Toallas", [
        "Cómo usar esta plantilla:",
        "▸ 4 checklists: apertura, cierre, mantenimiento semanal y servicio de toallas.",
        "▸ Incluye control de químicos, seguridad y cumplimiento normativo.",
        "▸ Obligatorio registrar pH y cloro en libro de mantenimiento.",
        "",
        "Personalización:",
        "▸ Ajusta según normativa local de piscinas (varía por comunidad autónoma).",
        "▸ Añade protocolos específicos si tienes piscina climatizada/cubierta.",
    ])
    create_task_sheet(wb, "Apertura Piscina", PISCINA_COLOR,
                      "Apertura de Piscina — Checklist Diario", [
        ("Control de agua y seguridad", [
            ("Medir cloro y pH del agua y registrar en libro obligatorio", "Piscina"),
            ("Pasar limpiafondos automático o manual y recoger hojas con red", "Piscina"),
            ("Verificar nivel de agua y rellenar si necesario", "Piscina"),
            ("Limpiar bordes del vaso, playa y zona de duchas", "Piscina"),
        ]),
        ("Montaje de zona", [
            ("Distribuir tumbonas y sombrillas según layout establecido", "Piscina"),
            ("Colocar señalización de profundidad y normas de uso visibles", "Piscina"),
            ("Verificar salvavidas, botiquín y material de rescate en su sitio", "Piscina"),
            ("Limpiar duchas y lavapiés y verificar que funcionan", "Piscina"),
        ]),
    ])
    create_task_sheet(wb, "Cierre Piscina", PISCINA_COLOR,
                      "Cierre de Piscina — Checklist Diario", [
        ("Recogida y cierre", [
            ("Recoger todas las toallas abandonadas y enviar a lavandería", "Piscina"),
            ("Apilar tumbonas y cubrir con lonas si procede", "Piscina"),
            ("Segunda medición de cloro y pH (registrar en libro)", "Piscina"),
            ("Verificar que la piscina queda vacía de bañistas", "Piscina"),
            ("Apagar iluminación de la zona de piscina", "Piscina"),
            ("Recoger material de juego/flotadores y almacenar", "Piscina"),
            ("Cerrar acceso a la zona de piscina según protocolo", "Piscina"),
        ]),
    ])
    create_task_sheet(wb, "Mant. Semanal", PISCINA_COLOR,
                      "Mantenimiento Semanal de Piscina", [
        ("Mantenimiento técnico", [
            ("Realizar retrolavado del sistema de filtración", "Piscina"),
            ("Verificar bombas de circulación y dosificación automática", "Piscina"),
            ("Comprobar stock de productos químicos en depósito", "Piscina"),
            ("Limpiar skimmers, rejillas de desbordamiento y sumideros", "Piscina"),
            ("Verificar escaleras, pasamanos y juntas del vaso", "Piscina"),
            ("Comprobar iluminación subacuática y exterior", "Piscina"),
            ("Actualizar libro de mantenimiento obligatorio con todas las actuaciones", "Piscina"),
        ]),
    ])
    create_task_sheet(wb, "Servicio Toallas", PISCINA_COLOR,
                      "Servicio de Toallas de Piscina", [
        ("Operativa de toallas", [
            ("Montar puesto de entrega de toallas con stock suficiente", "Piscina"),
            ("Sistema de entrega: por habitación, pulsera o tarjeta de toalla", "Piscina"),
            ("Reponer stock a mediodía (segundo envío desde lavandería)", "Piscina"),
            ("Recoger toallas abandonadas en tumbonas cada hora", "Piscina"),
            ("Conteo final de toallas y envío a lavandería con albarán", "Piscina"),
            ("Registrar mermas y pérdidas (cargo a habitación si procede)", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "11-piscina.xlsx"))


# ═══════════════════════════════════════════════════════════
# 12 — TERRAZA
# ═══════════════════════════════════════════════════════════
def gen_12():
    wb = Workbook()
    add_instructions_sheet(wb, "12 · Terraza — Montaje Temporada y Servicio Diario", [
        "Cómo usar esta plantilla:",
        "▸ 2 checklists: montaje de temporada y servicio diario de terraza.",
        "▸ Ideal para hoteles con terraza-restaurante, rooftop o jardín.",
        "▸ Incluye coordinación con F&B para carta de terraza.",
        "",
        "Personalización:",
        "▸ Ajusta según el tipo de terraza de tu hotel.",
        "▸ Añade protocolos de climatología adversa.",
    ])
    create_task_sheet(wb, "Montaje Temporada", TERRAZA_COLOR,
                      "Montaje de Terraza — Inicio de Temporada", [
        ("Preparación de temporada", [
            ("Sacar mobiliario de almacén: limpiar mesas, sillas, bancos", "Terraza"),
            ("Verificar estado de cojines, telas y parasoles (reemplazar dañados)", "Terraza"),
            ("Montar terraza según plano de distribución aprobado", "Terraza"),
            ("Instalar y verificar iluminación exterior (guirnaldas, focos, velas)", "Terraza"),
            ("Comprobar tomas de corriente para equipos (TPV, música, calefactores)", "Terraza"),
            ("Colocar jardineras, plantas y elementos decorativos", "Terraza"),
            ("Coordinar con F&B: carta de terraza, stock, personal de servicio", "Terraza"),
        ]),
    ])
    create_task_sheet(wb, "Servicio Diario", TERRAZA_COLOR,
                      "Terraza — Servicio Diario", [
        ("Apertura y mantenimiento diario", [
            ("Limpiar mesas y sillas a primera hora de la mañana", "Terraza"),
            ("Abrir parasoles y anclar según previsión de viento", "Terraza"),
            ("Colocar ceniceros en zona de fumadores", "Terraza"),
            ("Colocar cartas, condimentos y servilleteros en cada mesa", "Terraza"),
            ("Barrer y baldear suelo de la terraza", "Terraza"),
        ]),
        ("Durante el servicio", [
            ("Mantener mesas limpias: recoger platos y limpiar inmediatamente", "Terraza"),
            ("Vigilar previsión meteorológica: recoger cojines si llueve", "Terraza"),
            ("Reponer condimentos, servilletas y cartas según necesidad", "Terraza"),
            ("Verificar que iluminación exterior funciona al anochecer", "Terraza"),
            ("Recoger terraza completa al cierre: cojines, cartas, decoración", "Terraza"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "12-terraza.xlsx"))


# ═══════════════════════════════════════════════════════════
# 13 — MANTENIMIENTO
# ═══════════════════════════════════════════════════════════
def gen_13():
    wb = Workbook()
    add_instructions_sheet(wb, "13 · Mantenimiento — 6 Áreas Técnicas", [
        "Cómo usar esta plantilla:",
        "▸ 6 checklists: ronda diaria, semanal, mensual, HVAC, fontanería, electricidad.",
        "▸ Cubre mantenimiento preventivo y correctivo del hotel.",
        "▸ Esencial para cumplimiento normativo y eficiencia energética.",
        "",
        "Personalización:",
        "▸ Ajusta según la antigüedad y equipamiento de tu hotel.",
        "▸ Añade equipos específicos que tenga tu instalación.",
    ])
    create_task_sheet(wb, "Ronda Diaria", MANTENIMIENTO_COLOR,
                      "Mantenimiento — Ronda Diaria", [
        ("Ronda técnica diaria (07:00-08:00)", [
            ("Verificar calderas y producción de ACS (agua caliente sanitaria)", "Mantenimiento"),
            ("Revisar cuadros eléctricos principales: sin alarmas ni disparos", "Mantenimiento"),
            ("Comprobar grupo electrógeno: nivel combustible y estado stand-by", "Mantenimiento"),
            ("Verificar temperaturas HVAC en zonas comunes y habitaciones test", "Mantenimiento"),
            ("Comprobar funcionamiento de bombas de agua (presión, ruidos)", "Mantenimiento"),
            ("Ronda visual por áreas comunes: lobby, pasillos, fachada", "Mantenimiento"),
            ("Revisar parte de averías y priorizar reparaciones del día", "Mantenimiento"),
        ]),
    ])
    create_task_sheet(wb, "Semanal", MANTENIMIENTO_COLOR,
                      "Mantenimiento — Tareas Semanales", [
        ("Revisiones semanales", [
            ("Limpiar filtros de AC en zonas comunes (restaurante, lobby, salones)", "Mantenimiento"),
            ("Verificar extintores y BIEs: precintos, presión, accesibilidad", "Mantenimiento"),
            ("Test de detectores de incendio (muestra rotativa por planta)", "Mantenimiento"),
            ("Verificar funcionamiento de ascensores: puertas, botones, iluminación", "Mantenimiento"),
            ("Engrasar bisagras de puertas de emergencia y verificar cierre", "Mantenimiento"),
            ("Verificar bombas de piscina y spa: presión, ruidos, fugas", "Mantenimiento"),
            ("Comprobar alumbrado de emergencia (test de encendido)", "Mantenimiento"),
            ("Lectura de contadores: agua, gas, electricidad (registrar consumos)", "Mantenimiento"),
        ]),
    ])
    create_task_sheet(wb, "Mensual", MANTENIMIENTO_COLOR,
                      "Mantenimiento — Tareas Mensuales", [
        ("Revisiones mensuales", [
            ("Revisar equipos de cocina industrial y campanas extractoras", "Mantenimiento"),
            ("Inspeccionar cubiertas, canalones y bajantes", "Mantenimiento"),
            ("Revisar estado de pintura exterior e interior (zonas de uso)", "Mantenimiento"),
            ("Verificar puertas cortafuegos: cierre, retenedores, señalización", "Mantenimiento"),
            ("Probar grupo electrógeno bajo carga (test mensual obligatorio)", "Mantenimiento"),
            ("Revisar fontanería: desagües, sifones, posibles fugas", "Mantenimiento"),
            ("Inspeccionar mobiliario de habitaciones: cajones, bisagras, tiradores", "Mantenimiento"),
            ("Actualizar registro de mantenimiento y planificar mes siguiente", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "HVAC", MANTENIMIENTO_COLOR,
                      "HVAC — Climatización y Ventilación", [
        ("Control de climatización", [
            ("Verificar consignas de temperatura por zona (habitaciones, lobby, salones)", "HVAC"),
            ("Comprobar nivel de refrigerante en unidades de producción", "HVAC"),
            ("Limpiar filtros de fancoils trimestral (rotar por plantas)", "HVAC"),
            ("Revisar termostatos: incidencias de habitaciones, calibración", "HVAC"),
            ("Verificar ventilación y renovación de aire en zonas interiores", "HVAC"),
            ("Registrar consumo energético y comparar con período anterior", "HVAC"),
        ]),
    ])
    create_task_sheet(wb, "Fontanería", MANTENIMIENTO_COLOR,
                      "Fontanería — Control y Prevención", [
        ("Revisiones de fontanería", [
            ("Verificar presión de agua en puntos altos del edificio", "Fontanería"),
            ("Controlar temperatura de ACS para prevención de legionella (>50°C)", "Fontanería"),
            ("Inspeccionar cuartos técnicos: fugas, humedad, corrosión", "Fontanería"),
            ("Realizar desatascos preventivos en zonas de alto uso", "Fontanería"),
            ("Verificar cisternas: consumo fantasma, mecanismo, flotador", "Fontanería"),
            ("Revisar grifería de habitaciones: goteos, cal, funcionamiento", "Fontanería"),
            ("Verificar radiadores/calefacción: purgado, fugas, termostatos", "Fontanería"),
        ]),
    ])
    create_task_sheet(wb, "Electricidad", MANTENIMIENTO_COLOR,
                      "Electricidad — Control y Seguridad", [
        ("Revisiones eléctricas", [
            ("Verificar cuadros secundarios por planta: sin disparos ni alarmas", "Electricidad"),
            ("Comprobar alumbrado de emergencia y señalización evacuación", "Electricidad"),
            ("Revisar luminarias LED: reemplazar fundidas, verificar reguladores", "Electricidad"),
            ("Verificar tomas de corriente en habitaciones y zonas comunes", "Electricidad"),
            ("Comprobar cerraduras electrónicas: baterías, programación, backup", "Electricidad"),
            ("Verificar sistema de megafonía y música ambiental", "Electricidad"),
            ("Comprobar cargadores de vehículos eléctricos si existen", "Electricidad"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "13-mantenimiento.xlsx"))


# ═══════════════════════════════════════════════════════════
# 14 — ADMINISTRACIÓN
# ═══════════════════════════════════════════════════════════
def gen_14():
    wb = Workbook()
    add_instructions_sheet(wb, "14 · Administración — Revenue, Reservas, Contabilidad, RRHH", [
        "Cómo usar esta plantilla:",
        "▸ 4 checklists para la gestión administrativa diaria del hotel.",
        "▸ Revenue management, reservas, contabilidad y RRHH operativo.",
        "▸ Ideal para directores de hotel, revenue managers y back office.",
        "",
        "Personalización:",
        "▸ Ajusta KPIs y procesos a tu tipo de hotel.",
        "▸ Añade los informes específicos de tu PMS/channel manager.",
    ])
    create_task_sheet(wb, "Revenue Management", ADMIN_OFICINA_COLOR,
                      "Revenue Management — Control Diario", [
        ("Análisis diario", [
            ("Revisar ocupación y pickup vs. presupuesto (hoy, semana, mes)", "Revenue"),
            ("Verificar paridad de precios en OTAs: Booking, Expedia, canal directo", "Revenue"),
            ("Ajustar pricing dinámico según demanda y competencia", "Revenue"),
            ("Gestionar restricciones: estancia mínima, CTA, CTD según ocupación", "Revenue"),
            ("Revisar cancelaciones y no-shows del día anterior", "Revenue"),
            ("Analizar estado de grupos y eventos confirmados/tentative", "Revenue"),
            ("Preparar forecast semanal: ocupación, ADR, RevPAR proyectado", "Revenue"),
        ]),
    ])
    create_task_sheet(wb, "Reservas", ADMIN_OFICINA_COLOR,
                      "Departamento de Reservas — Operativa Diaria", [
        ("Gestión de reservas", [
            ("Procesar reservas nuevas: directas, OTAs, agencias, corporativas", "Reservas"),
            ("Confirmar garantía de reservas (tarjeta de crédito, prepago)", "Reservas"),
            ("Gestionar modificaciones y cancelaciones del día", "Reservas"),
            ("Room assignment: pre-asignar habitaciones según preferencias y tipo", "Reservas"),
            ("Comunicar peticiones especiales a departamentos (cuna, cama extra, dieta)", "Reservas"),
            ("Gestionar situación de overbooking y soluciones disponibles", "Reservas"),
            ("Enviar confirmaciones pre-arrival con información del hotel", "Reservas"),
        ]),
    ])
    create_task_sheet(wb, "Contabilidad", ADMIN_OFICINA_COLOR,
                      "Contabilidad — Control Diario", [
        ("Cierre diario", [
            ("Cuadrar cajas de recepción, restaurantes y spa", "Contabilidad"),
            ("Verificar cargos a habitación correctamente imputados", "Contabilidad"),
            ("Procesar facturas de empresas y créditos", "Contabilidad"),
            ("Conciliar cobros con TPV (Visa, Mastercard, Amex)", "Contabilidad"),
            ("Generar informe de ingresos por departamento (P&L diario)", "Contabilidad"),
            ("Gestionar aprobación de facturas de proveedores", "Contabilidad"),
            ("Preparar remesas bancarias y verificar movimientos", "Contabilidad"),
        ]),
    ])
    create_task_sheet(wb, "RRHH Operativo", ADMIN_OFICINA_COLOR,
                      "RRHH Operativo — Gestión Diaria", [
        ("Control de personal", [
            ("Verificar cuadrante de turnos y gestionar ausencias/bajas", "RRHH"),
            ("Controlar fichaje de entrada/salida de todo el personal", "RRHH"),
            ("Gestionar solicitudes de vacaciones y permisos", "RRHH"),
            ("Coordinar formaciones programadas del mes", "RRHH"),
            ("Verificar uniformidad del personal en servicio", "RRHH"),
            ("Gestionar incidencias de nómina: horas extra, complementos, dietas", "RRHH"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "14-administracion.xlsx"))


# ═══════════════════════════════════════════════════════════
# 15 — SPA & WELLNESS
# ═══════════════════════════════════════════════════════════
def gen_15():
    wb = Workbook()
    add_instructions_sheet(wb, "15 · Spa & Wellness — Apertura, Cierre, Cabinas, Vestuarios", [
        "Cómo usar esta plantilla:",
        "▸ 4 checklists: apertura, cierre, entre tratamientos y vestuarios.",
        "▸ Cubre piscina climatizada, jacuzzi, sauna, vapor y cabinas.",
        "▸ Ideal para responsables de spa y terapeutas.",
        "",
        "Personalización:",
        "▸ Ajusta según las instalaciones de tu spa.",
        "▸ Añade protocolos de higiene específicos de tu marca.",
    ])
    create_task_sheet(wb, "Apertura Spa", SPA_COLOR,
                      "Apertura del Spa — Checklist Diario", [
        ("Puesta en marcha de instalaciones", [
            ("Verificar temperatura de piscina/jacuzzi (35-38°C) y ajustar", "Spa"),
            ("Encender sauna y baño de vapor (precalentar 30-45 min)", "Spa"),
            ("Preparar cabinas: sábanas limpias, aceites, velas, música ambiental", "Spa"),
            ("Verificar stock de productos de tratamiento (cremas, aceites, exfoliantes)", "Spa"),
            ("Revisar agenda del día y confirmar terapeutas asignados", "Spa"),
        ]),
        ("Preparación de zonas", [
            ("Limpiar y preparar vestuarios: duchas, taquillas, amenities", "Vestuarios"),
            ("Ajustar iluminación tenue y sistema de sonido ambiental", "Spa"),
            ("Preparar zona de relajación: agua, infusiones, fruta fresca", "Spa"),
        ]),
    ])
    create_task_sheet(wb, "Cierre Spa", SPA_COLOR,
                      "Cierre del Spa — Checklist Diario", [
        ("Verificación y apagado", [
            ("Verificar que todas las zonas están vacías de clientes", "Spa"),
            ("Apagar saunas, baño de vapor y jacuzzi", "Spa"),
            ("Enviar todas las toallas usadas a lavandería con albarán", "Spa"),
            ("Limpiar y desinfectar todas las cabinas de tratamiento", "Spa"),
            ("Limpieza a fondo de vestuarios: suelo, duchas, taquillas", "Vestuarios"),
        ]),
        ("Cierre administrativo", [
            ("Cuadrar caja: tratamientos realizados, cargos a habitación, ventas producto", "Admin"),
            ("Registrar incidencias del día y stock bajo de productos", "Admin"),
            ("Revisar agenda del día siguiente y confirmar disponibilidad", "Spa"),
        ]),
    ])
    create_task_sheet(wb, "Cabinas Tratamientos", SPA_COLOR,
                      "Cabinas — Protocolo Entre Tratamientos", [
        ("Reset de cabina (10-15 min entre clientes)", [
            ("Cambiar sábanas y toallas de la camilla completamente", "Spa"),
            ("Desinfectar todas las superficies: camilla, taburete, estanterías", "Spa"),
            ("Reponer productos necesarios para el siguiente tratamiento", "Spa"),
            ("Ventilar la cabina durante 5 minutos mínimo", "Spa"),
            ("Ajustar temperatura de cabina a 24-26°C", "Spa"),
            ("Preparar set-up específico del siguiente tratamiento (aceites, piedras, etc.)", "Spa"),
        ]),
    ])
    create_task_sheet(wb, "Vestuarios", SPA_COLOR,
                      "Vestuarios del Spa — Mantenimiento Continuo", [
        ("Limpieza y control (cada 2-3 horas)", [
            ("Fregar suelo antideslizante con producto desinfectante", "Vestuarios"),
            ("Desinfectar bancos, taquillas y manillas de puertas", "Vestuarios"),
            ("Reponer amenities: gel, champú, crema corporal, secadores", "Vestuarios"),
            ("Verificar que taquillas funcionan correctamente (llaves/códigos)", "Vestuarios"),
            ("Vaciar papeleras y recoger toallas usadas del suelo", "Vestuarios"),
            ("Verificar ventilación: sin malos olores, extracción funcionando", "Vestuarios"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "15-spa-wellness.xlsx"))


# ═══════════════════════════════════════════════════════════
# BONUS 01 — BRIEFING DIARIO F&B
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 01 · Briefing Diario F&B del Hotel", [
        "Cómo usar esta plantilla:",
        "▸ Plantilla para la reunión diaria del departamento F&B.",
        "▸ Incluye ocupación, VIPs, eventos del día, menú y equipo.",
        "▸ Rellena cada mañana y reparte a jefes de partida y sala.",
        "",
        "Personalización:",
        "▸ Adapta los outlets a los que tenga tu hotel.",
        "▸ Imprime una copia para cada jefe de sección.",
    ])
    create_task_sheet(wb, "Briefing", ADMIN_COLOR, "Briefing Diario F&B", [
        ("Datos del día", [
            ("Fecha: ___/___/___  Día: ___  Ocupación: ___% (___hab. de ___)", "Admin"),
            ("Check-ins hoy: ___  Check-outs hoy: ___  Grupos: ___", "Admin"),
            ("Previsión comensales desayuno: ___  Almuerzo: ___  Cena: ___", "Admin"),
            ("Room service previstos: ___  Minibar reposiciones: ___", "Admin"),
            ("Eventos/banquetes hoy: _______________ Salón: ___  Pax: ___", "Admin"),
        ]),
        ("VIPs y atención especial", [
            ("VIP 1: Hab.___ Nombre:_______________ Atención:_______________", "Admin"),
            ("VIP 2: Hab.___ Nombre:_______________ Atención:_______________", "Admin"),
            ("Alérgenos críticos: Hab.___ Tipo:_______________ Protocolo:___", "Admin"),
            ("Dietas especiales (vegano, kosher, halal): _______________", "Admin"),
            ("Celebraciones: cumpleaños/aniversario — Hab.___ Detalle:___", "Admin"),
        ]),
        ("Menú y producto del día", [
            ("Sugerencia del chef almuerzo: _______________", "Cocina"),
            ("Sugerencia del chef cena: _______________", "Cocina"),
            ("Productos agotados o no disponibles: _______________", "Cocina"),
            ("Show cooking del día: _______________", "Cocina"),
            ("Vino recomendado: _______________", "Sala"),
        ]),
        ("Equipo del día", [
            ("Jefe de cocina en turno: _______________", "Admin"),
            ("Maître en turno: _______________", "Admin"),
            ("Room service responsable: _______________", "Admin"),
            ("Banquetes responsable: _______________", "Admin"),
            ("Personal extra/eventual: ___ personas", "Admin"),
            ("Incidencias pendientes del día anterior: _______________", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))


# ═══════════════════════════════════════════════════════════
# BONUS 02 — CALENDARIO ANUAL HOTELERO
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 02 · Calendario Anual F&B Hotelero", [
        "Cómo usar esta plantilla:",
        "▸ Las 20 fechas y temporadas clave para F&B en hoteles.",
        "▸ Incluye temporadas alta/baja, festividades y eventos especiales.",
        "▸ Planifica menús especiales, personal extra y compras anticipadas.",
        "",
        "Personalización:",
        "▸ Añade fechas locales (fiestas patronales, ferias, congresos).",
        "▸ Marca las fechas que afectan a tu hotel según su tipo.",
    ])

    ws = wb.create_sheet(title="Calendario")
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Calendario Anual F&B — Hotel"
    ws["A1"].font = title_font

    ws.merge_cells("A2:F2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — 20 fechas clave para F&B hotelero"
    ws["A2"].font = subtitle_font

    headers = ["#", "Fecha / Período", "Impacto", "Preparación F&B", "Antelación", "Notas"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    events = [
        ("Enero — Reyes Magos", "Desayuno especial", "Roscón gigante en buffet, menú infantil especial", "3 semanas"),
        ("Febrero — San Valentín", "Cena temática", "Menú degustación para parejas, decoración especial", "4 semanas"),
        ("Marzo — Semana Santa", "Ocupación alta", "Menús tradicionales, refuerzo de personal, buffet ampliado", "6 semanas"),
        ("Abril — Temporada media", "Grupos/congresos", "Preparar paquetes de coffee break y almuerzos de trabajo", "2 meses"),
        ("Mayo — Bodas (inicio)", "Banquetes pico", "Reservar personal extra, confirmar proveedores de boda", "6-12 meses"),
        ("Junio — Temporada alta", "Máx. ocupación", "Ampliar horarios buffet, más show cooking, pool bar", "3 meses"),
        ("Jul-Ago — Verano pleno", "All-inclusive/resort", "Menús refrescantes, helados, chiringuito, BBQ nocturna", "3 meses"),
        ("Septiembre — Vuelta congresos", "Corporativo fuerte", "Kick-offs, convenciones. Preparar paquetes de banquetes", "2 meses"),
        ("Octubre — Temporada de caza", "Menú especial", "Incorporar caza y setas a carta y buffet", "4 semanas"),
        ("Octubre — Halloween", "Evento temático", "Buffet temático, cocktails especiales, decoración", "4 semanas"),
        ("Nov — Black Friday gastro", "Promoción", "Paquetes gastro: cena+alojamiento, brunch dominical", "4 semanas"),
        ("Dic 1-20 — Cenas empresa", "Banquetes max", "Máxima demanda de banquetes corporativos del año", "3 meses"),
        ("24 Dic — Nochebuena", "Cena especial", "Menú premium de Nochebuena. Reservas con antelación", "2 meses"),
        ("25 Dic — Navidad", "Almuerzo familiar", "Buffet de Navidad ampliado. Turrones y dulces", "2 meses"),
        ("31 Dic — Nochevieja", "Gala + cotillón", "Cena de gala + cotillón + barra libre. Personal máximo", "3 meses"),
        ("Todo el año — Brunch dominical", "Outlet fijo", "Brunch premium con champagne, huevos benedict, show", "Semanal"),
        ("Todo el año — Amenities VIP", "Room service", "Welcome drinks, fruit baskets, chocolate turndown", "Stock mensual"),
        ("Todo el año — Grupos MICE", "Banquetes", "Coffee breaks, almuerzos de trabajo, cenas de gala", "1-3 meses"),
        ("Festivos locales", "Variable", "Menú especial según festividad local. Consultar calendario", "4-6 semanas"),
        ("Inspección sanidad", "Obligatorio", "Preparar toda documentación APPCC, temperaturas, formación", "Permanente"),
    ]

    impacto_colors = {
        "Desayuno especial": "FFF3E0",
        "Cena temática": "FCE4EC",
        "Ocupación alta": "E8F5E9",
        "Grupos/congresos": "E3F2FD",
        "Banquetes pico": "F3E5F5",
        "Máx. ocupación": "E8F5E9",
        "All-inclusive/resort": "E0F2F1",
        "Corporativo fuerte": "E3F2FD",
        "Menú especial": "FFF3E0",
        "Evento temático": "F3E5F5",
        "Promoción": "FFF8E1",
        "Banquetes max": "F3E5F5",
        "Cena especial": "FCE4EC",
        "Almuerzo familiar": "FFF3E0",
        "Gala + cotillón": "F3E5F5",
        "Outlet fijo": "E0F2F1",
        "Room service": "E3F2FD",
        "Banquetes": "F3E5F5",
        "Variable": LIGHT_GRAY,
        "Obligatorio": "FFCDD2",
    }

    for idx, (date, impacto, prep, antelacion) in enumerate(events, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx).font = data_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=date).font = bold_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        color = impacto_colors.get(impacto, LIGHT_GRAY)
        ws.cell(row=row, column=3, value=impacto).font = data_font
        ws.cell(row=row, column=3).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=prep).font = data_font
        ws.cell(row=row, column=4).alignment = left_align
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=antelacion).font = bold_font
        ws.cell(row=row, column=5).alignment = center_align
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).fill = input_fill

    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx"))


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Generando Kit de Tareas: Hotel Integral (15 archivos + 2 BONUS)...")
    print()

    gen_01()
    print("  ✓ 01 — Buffet Desayuno (Apertura + Cierre)")
    gen_02()
    print("  ✓ 02 — Buffet Comida/Cena (Almuerzo + Cena)")
    gen_03()
    print("  ✓ 03 — Restaurante À la Carte (Apertura + Cierre)")
    gen_04()
    print("  ✓ 04 — F&B Outlets (Pool Bar + Lobby Bar + Snack Bar)")
    gen_05()
    print("  ✓ 05 — Room Service + Minibar")
    gen_06()
    print("  ✓ 06 — Banquetes y Eventos")
    gen_07()
    print("  ✓ 07 — Recepción (3 Turnos + Check-in/out)")
    gen_08()
    print("  ✓ 08 — Guest Services (Conserjería + Guest Experience)")
    gen_09()
    print("  ✓ 09 — Housekeeping (5 protocolos)")
    gen_10()
    print("  ✓ 10 — Áreas Públicas (Lobby + Pasillos + Baños + Parking)")
    gen_11()
    print("  ✓ 11 — Piscina (Apertura + Cierre + Semanal + Toallas)")
    gen_12()
    print("  ✓ 12 — Terraza (Montaje Temporada + Servicio Diario)")
    gen_13()
    print("  ✓ 13 — Mantenimiento (6 áreas técnicas)")
    gen_14()
    print("  ✓ 14 — Administración (Revenue + Reservas + Contabilidad + RRHH)")
    gen_15()
    print("  ✓ 15 — Spa & Wellness (Apertura + Cierre + Cabinas + Vestuarios)")

    print()
    gen_bonus_01()
    print("  ✓ BONUS 01 — Briefing Diario F&B")
    gen_bonus_02()
    print("  ✓ BONUS 02 — Calendario Anual F&B Hotelero")

    print(f"\n✅ 17 archivos generados en {OUTPUT_DIR}")
    print("   (15 plantillas principales + 2 BONUS)")

    # Summary
    sheet_count = {
        "01": 2, "02": 2, "03": 2, "04": 3, "05": 2, "06": 1,
        "07": 4, "08": 2, "09": 5, "10": 4, "11": 4, "12": 2,
        "13": 6, "14": 4, "15": 4,
    }
    total = sum(sheet_count.values())
    print(f"   Total: {total} plantillas de tareas en {len(sheet_count)} archivos")
