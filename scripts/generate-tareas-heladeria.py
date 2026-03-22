#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Heladería Artesanal (9 archivos).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-heladeria"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Zone colors for heladería
OBRADOR_COLOR = "FFF3E0"      # Orange light - production
VITRINA_COLOR = "E3F2FD"      # Blue light - display/service
CAMARA_COLOR = "E0F7FA"       # Cyan light - cold storage
MOSTRADOR_COLOR = "F3E5F5"    # Purple light - counter
LIMPIEZA_COLOR = "EFEBE9"     # Brown light - cleaning
ADMIN_COLOR = "FFF8E1"        # Amber light - admin
ALMACEN_COLOR = "E8F5E9"      # Green light - storage
TERRAZA_COLOR = "F1F8E9"      # Light green - terrace

ZONE_COLORS = {
    "Obrador": OBRADOR_COLOR,
    "Pasteurizador": OBRADOR_COLOR,
    "Mantecadora": OBRADOR_COLOR,
    "Producción": OBRADOR_COLOR,
    "Vitrina": VITRINA_COLOR,
    "Cámara": CAMARA_COLOR,
    "Maduración": CAMARA_COLOR,
    "Abatidor": CAMARA_COLOR,
    "Mostrador": MOSTRADOR_COLOR,
    "Caja": MOSTRADOR_COLOR,
    "Servicio": MOSTRADOR_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Gestión": ADMIN_COLOR,
    "Almacén": ALMACEN_COLOR,
    "Terraza": TERRAZA_COLOR,
    "General": LIGHT_GRAY,
    "Equipo": ADMIN_COLOR,
    "Exterior": TERRAZA_COLOR,
    "Logística": ALMACEN_COLOR,
    "RRSS": ADMIN_COLOR,
}

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Heladería Artesanal"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for section_title, tasks in sections:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for idx, (task, zone) in enumerate(tasks, 1):
            zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")

            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=zone).font = data_font
            ws.cell(row=row, column=3).fill = zone_fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align

            check_cell = ws.cell(row=row, column=5)
            check_cell.font = checkbox_font
            check_cell.alignment = center_align
            check_cell.border = thin_border
            dv.add(check_cell)

            ws.cell(row=row, column=6).fill = input_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


def create_blank_template_sheet(wb, sheet_name, tab_color, title, col_headers):
    """Create a blank customizable sheet with headers and empty rows."""
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Heladería Artesanal"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r in range(5, 30):
        for c in range(1, len(col_headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c == 1:
                cell.alignment = center_align
            elif c == 5:
                cell.font = checkbox_font
                cell.alignment = center_align
                dv.add(cell)
            elif c in (4, 6):
                cell.fill = input_fill
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.merge_cells(f"A31:G31")
    ws[f"A31"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A31"].font = small_font


def create_calendar_sheet(wb, sheet_name, tab_color, title, months_data):
    """Create a calendar-style sheet with monthly rows."""
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:F2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Heladería Artesanal"
    ws["A2"].font = subtitle_font

    row = 4
    cal_headers = ["Mes", "Acciones clave / Fechas señaladas", "Sabores destacados", "Temporada", "✓", "Notas"]
    for col_idx, h in enumerate(cal_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    ws.add_data_validation(dv)

    for month, actions, flavors, season in months_data:
        ws.cell(row=row, column=1, value=month).font = bold_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=actions).font = data_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=flavors).font = data_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = thin_border

        season_color = OBRADOR_COLOR if season == "Alta" else CAMARA_COLOR if season == "Baja" else VITRINA_COLOR
        ws.cell(row=row, column=4, value=season).font = data_font
        ws.cell(row=row, column=4).fill = PatternFill(start_color=season_color, end_color=season_color, fill_type="solid")
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border

        check = ws.cell(row=row, column=5)
        check.font = checkbox_font
        check.alignment = center_align
        check.border = thin_border
        dv.add(check)

        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = left_align

        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


# ═══════════════════════════════════════════════════════════
# 01 — APERTURA Y CIERRE DE HELADERÍA
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Apertura y Cierre de Heladería", [
        "Cómo usar esta plantilla:",
        "▸ Checklist completo para la apertura y cierre diario de tu heladería.",
        "▸ Incluye encendido de vitrinas, cámaras, montaje de servicio y cierre.",
        "▸ Marca ✓ cada tarea completada. Adapta horarios a tu negocio.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables: responsable, hora, notas.",
        "▸ Ajusta temperaturas según tu equipamiento y normativa local.",
        "▸ Imprime en A4 y deja en obrador y mostrador.",
    ])
    create_task_sheet(wb, "Apertura", OBRADOR_COLOR, "Apertura de Heladería", [
        ("Encendido y verificación de equipos", [
            ("Encender vitrinas de exposición y verificar temperatura -14°C/-12°C", "Vitrina"),
            ("Verificar cámaras de conservación a -18°C", "Cámara"),
            ("Verificar cámara de maduración a 2-4°C", "Maduración"),
            ("Comprobar abatidor de temperatura operativo", "Abatidor"),
            ("Encender iluminación de vitrinas y sala", "Vitrina"),
            ("Registrar temperaturas en hoja de control APPCC", "Admin"),
        ]),
        ("Montaje de servicio", [
            ("Montar toppings frescos: fruta, frutos secos, virutas de chocolate", "Mostrador"),
            ("Preparar salsas: chocolate, caramelo, frutos rojos, dulce de leche", "Mostrador"),
            ("Reponer cucuruchos, tarrinas, cucharillas y servilletas", "Mostrador"),
            ("Verificar stock de conos artesanales y barquillos", "Mostrador"),
            ("Limpiar cristales de vitrina (interior y exterior)", "Limpieza"),
            ("Verificar aspecto visual de cubetas en vitrina (sin cristales de hielo)", "Vitrina"),
        ]),
        ("Apertura al público", [
            ("Encender TPV / caja registradora y verificar fondo de caja", "Caja"),
            ("Verificar datáfono y terminal de pagos operativo", "Caja"),
            ("Colocar pizarra/carta exterior con sabores del día", "Mostrador"),
            ("Verificar limpieza de zona de terraza/exterior si procede", "Terraza"),
            ("Abrir puerta y encender rótulo exterior", "General"),
        ]),
    ])
    create_task_sheet(wb, "Cierre", OBRADOR_COLOR, "Cierre de Heladería", [
        ("Conservación de producto", [
            ("Tapar todas las cubetas con film/tapa hermética", "Vitrina"),
            ("Pasar vitrinas a modo conservación -18°C", "Vitrina"),
            ("Registrar mermas del día (sabores agotados, producto desechado)", "Admin"),
            ("Etiquetar cubetas con fecha de caducidad actualizada", "Vitrina"),
            ("Verificar stock de cubetas para producción del día siguiente", "Cámara"),
        ]),
        ("Limpieza de equipos y zona de servicio", [
            ("Limpiar y desinfectar vitrinas (superficies exteriores)", "Limpieza"),
            ("Vaciar y limpiar dispensadores de salsas", "Limpieza"),
            ("Limpiar y desinfectar porcionadores/espátulas de servicio", "Limpieza"),
            ("Limpiar mostrador, báscula y zona de cobro", "Limpieza"),
            ("Vaciar papeleras y reponer bolsas", "Limpieza"),
            ("Fregar suelos de sala, obrador y almacén", "Limpieza"),
            ("Recoger y limpiar zona de terraza si procede", "Terraza"),
        ]),
        ("Cierre administrativo", [
            ("Cuadre de caja: efectivo, tarjeta, otros medios de pago", "Caja"),
            ("Registrar ventas del día por sabor (top sellers)", "Admin"),
            ("Anotar incidencias del turno en libro de novedades", "Admin"),
            ("Apagar iluminación, rótulo exterior y equipos no esenciales", "General"),
            ("Cerrar y verificar puertas y alarma", "General"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx"))
    print("  ✓ 01-apertura-cierre.xlsx")


# ═══════════════════════════════════════════════════════════
# 02 — PARTIDAS DE PRODUCCIÓN
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Partidas de Producción", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para las dos fases clave de producción de helado artesanal.",
        "▸ Pasteurización: mezcla, tratamiento térmico, enfriamiento y maduración.",
        "▸ Mantecación: texturización, extracción, abatimiento y almacenaje.",
        "",
        "Personalización:",
        "▸ Ajusta temperaturas y tiempos a tu maquinaria y recetas.",
        "▸ Usa las notas para registrar lotes, rendimientos y overrun.",
        "▸ Imprime y deja en el obrador junto a las fichas técnicas.",
    ])
    create_task_sheet(wb, "Pasteurización", OBRADOR_COLOR, "Partida de Pasteurización", [
        ("Preparación de la mezcla", [
            ("Pesar todos los ingredientes según ficha técnica del sabor", "Obrador"),
            ("Mezclar ingredientes secos (azúcares, estabilizantes, leche en polvo)", "Obrador"),
            ("Añadir líquidos (leche, nata) al pasteurizador", "Pasteurizador"),
            ("Incorporar mezcla de secos sobre líquidos con batidora/turbo", "Pasteurizador"),
        ]),
        ("Pasteurización y enfriamiento", [
            ("Pasteurizar a 85°C durante 15 segundos (alta pasteurización)", "Pasteurizador"),
            ("Añadir grasas y emulsionantes al alcanzar 40-50°C en descenso", "Pasteurizador"),
            ("Enfriar rápidamente a 4°C en abatidor o ciclo de enfriamiento", "Abatidor"),
            ("Verificar temperatura final de la mezcla (<4°C)", "Obrador"),
        ]),
        ("Maduración", [
            ("Trasvasar mezcla a contenedor de maduración limpio", "Maduración"),
            ("Etiquetar: sabor, fecha, hora inicio maduración, lote", "Maduración"),
            ("Respetar tiempo de maduración: mínimo 4h, óptimo 8-12h a 2-4°C", "Maduración"),
            ("Limpiar y desinfectar pasteurizador tras cada uso", "Limpieza"),
            ("Registrar lote de producción en hoja de trazabilidad", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Mantecación", OBRADOR_COLOR, "Partida de Mantecación", [
        ("Preparación de mantecación", [
            ("Verificar que la mezcla ha completado tiempo de maduración", "Maduración"),
            ("Agitar suavemente la mezcla antes de verter en mantecadora", "Obrador"),
            ("Programar mantecadora: crema -7°C/-9°C | sorbete -10°C/-12°C", "Mantecadora"),
            ("Preparar cubetas pre-enfriadas en congelador (-20°C)", "Cámara"),
        ]),
        ("Mantecación y extracción", [
            ("Verter mezcla en mantecadora y comenzar ciclo", "Mantecadora"),
            ("Añadir variegatos, inclusiones o ripple en último minuto si procede", "Mantecadora"),
            ("Extraer helado en cubeta pre-enfriada con espátula", "Obrador"),
            ("Etiquetar cubeta: sabor, fecha producción, caducidad, lote", "Obrador"),
            ("Pesar cubeta y registrar rendimiento / overrun", "Obrador"),
        ]),
        ("Abatimiento y almacenaje", [
            ("Abatir cubeta a -18°C en abatidor (ciclo de congelación)", "Abatidor"),
            ("Trasladar cubeta a cámara de conservación -18°C/-20°C", "Cámara"),
            ("Registrar producción del día: sabor, kg, lote, destino", "Admin"),
            ("Limpiar mantecadora: desmontar cuchillas, junta, cilindro", "Limpieza"),
            ("Desinfectar superficies de obrador y utensilios utilizados", "Limpieza"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "02-partidas-produccion.xlsx"))
    print("  ✓ 02-partidas-produccion.xlsx")


# ═══════════════════════════════════════════════════════════
# 03 — TAREAS DEL MANAGER / ENCARGADO
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Tareas del Manager / Encargado", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para la gestión diaria, semanal y mensual de la heladería.",
        "▸ Cubre producción, ventas, equipo, compras y control financiero.",
        "▸ El encargado debe revisar estas tareas cada día/semana/mes.",
        "",
        "Personalización:",
        "▸ Ajusta frecuencias y responsabilidades según tu estructura.",
        "▸ Las celdas verdes son editables para personalizar.",
        "▸ Usa las notas para registrar KPIs y observaciones.",
    ])
    create_task_sheet(wb, "Tareas Diarias", ADMIN_COLOR, "Tareas Diarias del Manager", [
        ("Planificación y producción", [
            ("Revisar plan de producción del día (sabores y cantidades)", "Obrador"),
            ("Verificar stock de materias primas clave (leche, nata, azúcar, fruta)", "Almacén"),
            ("Revisar registros de temperatura APPCC del turno anterior", "Admin"),
            ("Comprobar mezclas en maduración listas para manteccar", "Maduración"),
        ]),
        ("Ventas y operaciones", [
            ("Revisar ventas del día anterior por sabor (identificar top sellers)", "Admin"),
            ("Asignar equipo a posiciones: obrador, mostrador, terraza", "Equipo"),
            ("Verificar pedidos de proveedores pendientes de recepción", "Almacén"),
            ("Gestionar incidencias del turno anterior (libro de novedades)", "Admin"),
        ]),
        ("Servicio y calidad", [
            ("Supervisar presentación de vitrina (atractivo visual, orden)", "Vitrina"),
            ("Verificar rotación de sabores en vitrina (FIFO)", "Vitrina"),
            ("Probar sabor del día / nueva receta si procede", "Obrador"),
            ("Atender reseñas online y responder comentarios", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Tareas Semanales", ADMIN_COLOR, "Tareas Semanales del Manager", [
        ("Análisis y control", [
            ("Analizar ventas por sabor de la semana (ranking, tendencias)", "Admin"),
            ("Calcular food cost semanal: objetivo <28-32%", "Admin"),
            ("Revisar cumplimiento de checklists de apertura/cierre", "Admin"),
            ("Reunión rápida con equipo: feedback, mejoras, planning", "Equipo"),
        ]),
        ("Compras y carta", [
            ("Realizar pedido semanal de materias primas", "Almacén"),
            ("Planificar sabores de la semana siguiente (rotación, estacionalidad)", "Obrador"),
            ("Actualizar carta/pizarra con sabores disponibles", "Mostrador"),
            ("Revisar stock de envases, cucuruchos, tarrinas y complementos", "Almacén"),
        ]),
        ("Mantenimiento", [
            ("Verificar estado de maquinaria: mantecadora, pasteurizador, vitrinas", "Obrador"),
            ("Supervisar limpieza profunda semanal de equipos", "Limpieza"),
            ("Revisar caducidades en almacén y cámaras (FIFO)", "Almacén"),
        ]),
    ])
    create_task_sheet(wb, "Tareas Mensuales", ADMIN_COLOR, "Tareas Mensuales del Manager", [
        ("Control financiero", [
            ("Elaborar P&L mensual: ventas, costes MP, personal, fijos", "Admin"),
            ("Realizar inventario completo valorado (MP, envases, producto)", "Almacén"),
            ("Revisar food cost mensual por familia de sabores", "Admin"),
            ("Evaluar rentabilidad por sabor (margen unitario)", "Admin"),
        ]),
        ("Equipo y formación", [
            ("Evaluación de rendimiento del equipo", "Equipo"),
            ("Planificar turnos y vacaciones del mes siguiente", "Equipo"),
            ("Sesión de formación: nuevas recetas, técnicas, atención al cliente", "Equipo"),
        ]),
        ("Producto y mantenimiento", [
            ("Mantenimiento preventivo de maquinaria (revisar juntas, filtros)", "Obrador"),
            ("Planificar carta estacional del mes siguiente", "Admin"),
            ("Actualizar fichas técnicas de sabores si hay cambios", "Admin"),
            ("Revisar y responder reseñas online del mes", "Admin"),
            ("Evaluar proveedores: calidad, precio, cumplimiento", "Admin"),
            ("Revisar documentación APPCC y registros del mes", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "03-tareas-manager.xlsx"))
    print("  ✓ 03-tareas-manager.xlsx")


# ═══════════════════════════════════════════════════════════
# 04 — TAREAS POR PERFIL PROFESIONAL
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Tareas por Perfil Profesional", [
        "Cómo usar esta plantilla:",
        "▸ Tres hojas con las tareas específicas de cada perfil.",
        "▸ Heladero/Obrador: producción, mantecación, limpieza de maquinaria.",
        "▸ Dependiente/Mostrador: atención al cliente, vitrina, caja.",
        "▸ Encargado de turno: supervisión, flujos, incidencias.",
        "",
        "Personalización:",
        "▸ Adapta según el tamaño de tu equipo y estructura.",
        "▸ Entrega a cada perfil su hoja correspondiente.",
    ])
    create_task_sheet(wb, "Heladero-Obrador", OBRADOR_COLOR, "Tareas del Heladero / Obrador", [
        ("Inicio de turno", [
            ("Revisar mezclas en maduración listas para manteccar", "Maduración"),
            ("Verificar estado y limpieza de pasteurizador y mantecadora", "Obrador"),
            ("Pesar ingredientes según fichas técnicas del plan de producción", "Obrador"),
            ("Preparar cubetas limpias y pre-enfriadas", "Cámara"),
        ]),
        ("Producción", [
            ("Preparar y pasteurizar mezclas según plan del día", "Pasteurizador"),
            ("Manteccar sabores según orden de prioridad (agotados primero)", "Mantecadora"),
            ("Añadir variegatos e inclusiones según ficha del sabor", "Mantecadora"),
            ("Rellenar vitrina con cubetas frescas (aspecto visual cuidado)", "Vitrina"),
            ("Registrar lotes de producción: sabor, kg, fecha, caducidad", "Admin"),
        ]),
        ("Cierre de turno", [
            ("Limpiar y desinfectar mantecadora (desmontar cuchillas y junta)", "Limpieza"),
            ("Limpiar pasteurizador si se ha utilizado", "Limpieza"),
            ("Limpiar obrador: superficies, balanzas, utensilios", "Limpieza"),
            ("Dejar preparadas mezclas en maduración para el día siguiente", "Maduración"),
            ("Comunicar incidencias y stock al encargado", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Dependiente-Mostrador", MOSTRADOR_COLOR, "Tareas del Dependiente / Mostrador", [
        ("Inicio de turno", [
            ("Verificar presentación de la vitrina (orden, limpieza, atractivo)", "Vitrina"),
            ("Montar toppings, salsas, cucuruchos y tarrinas en mostrador", "Mostrador"),
            ("Verificar operatividad de caja/TPV y datáfono", "Caja"),
            ("Revisar carta de sabores del día y alérgenos", "Mostrador"),
        ]),
        ("Durante el servicio", [
            ("Atender clientes: informar sabores, alérgenos, recomendar", "Servicio"),
            ("Servir helado con presentación cuidada (bola uniforme, topping)", "Servicio"),
            ("Cobrar correctamente y ofrecer tarjeta de fidelización", "Caja"),
            ("Mantener limpio mostrador, cristales de vitrina y zona de servicio", "Limpieza"),
            ("Avisar al obrador cuando un sabor esté agotado o bajo mínimo", "Servicio"),
            ("Gestionar colas y flujo de clientes en hora punta", "Servicio"),
        ]),
        ("Cierre de turno", [
            ("Cuadre de caja: efectivo, tarjeta y otros medios", "Caja"),
            ("Recoger toppings, salsas y guardar en refrigeración", "Mostrador"),
            ("Limpiar zona de mostrador y equipos de servicio", "Limpieza"),
            ("Comunicar al encargado incidencias y sabores más vendidos", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Encargado Turno", ADMIN_COLOR, "Tareas del Encargado de Turno", [
        ("Inicio de turno", [
            ("Supervisar apertura completa según checklist", "Admin"),
            ("Verificar equipo presente y asignar posiciones", "Equipo"),
            ("Revisar libro de novedades del turno anterior", "Admin"),
            ("Verificar temperaturas de vitrinas y cámaras (registro APPCC)", "Vitrina"),
        ]),
        ("Durante el turno", [
            ("Supervisar flujo de clientes y reforzar posiciones si necesario", "Servicio"),
            ("Gestionar reclamaciones y atención especial a clientes", "Servicio"),
            ("Monitorizar temperaturas de vitrina durante el servicio", "Vitrina"),
            ("Gestionar breaks del equipo sin afectar servicio", "Equipo"),
            ("Autorizar descuentos, cortesías o cambios de precio", "Admin"),
        ]),
        ("Cierre de turno", [
            ("Supervisar cierre completo según checklist", "Admin"),
            ("Registrar incidencias del turno en libro de novedades", "Admin"),
            ("Verificar cuadre de caja y firmar", "Caja"),
            ("Comunicar novedades relevantes al siguiente turno / manager", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "04-tareas-perfiles.xlsx"))
    print("  ✓ 04-tareas-perfiles.xlsx")


# ═══════════════════════════════════════════════════════════
# 05 — TAREAS SEMANALES Y MENSUALES
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Tareas Semanales y Mensuales", [
        "Cómo usar esta plantilla:",
        "▸ Limpieza profunda semanal, mantenimiento de equipos y control mensual.",
        "▸ Imprescindible para mantener calidad de producto y cumplir APPCC.",
        "▸ Asigna un día fijo de la semana para limpieza profunda.",
        "",
        "Personalización:",
        "▸ Adapta frecuencias según volumen de producción y equipamiento.",
        "▸ Las tareas mensuales se pueden distribuir a lo largo del mes.",
    ])
    create_task_sheet(wb, "Limpieza Semanal", LIMPIEZA_COLOR, "Limpieza Profunda Semanal", [
        ("Vitrinas y zona de servicio", [
            ("Desmontar y limpiar rejillas interiores de vitrinas de exposición", "Vitrina"),
            ("Limpiar interior completo de vitrinas (paredes, base, cristales)", "Vitrina"),
            ("Limpiar y desinfectar dispensadores de salsas y toppings", "Mostrador"),
            ("Desinfectar bandejas de cucuruchos, tarrinas y complementos", "Mostrador"),
        ]),
        ("Obrador y maquinaria", [
            ("Desmontar y limpiar piezas de mantecadora (cuchillas, junta, cilindro)", "Obrador"),
            ("Limpiar filtros de condensadores de equipos de frío", "Obrador"),
            ("Limpiar paredes y estantes de obrador", "Obrador"),
            ("Desinfectar contenedores de maduración", "Maduración"),
            ("Limpiar máquina de hielo si procede", "Obrador"),
        ]),
        ("Cámaras y almacén", [
            ("Limpiar interior de cámaras de conservación", "Cámara"),
            ("Revisar caducidades FIFO en cámaras y almacén", "Almacén"),
            ("Limpiar estantes y suelo de almacén", "Almacén"),
        ]),
    ])
    create_task_sheet(wb, "Gestión Semanal", ADMIN_COLOR, "Gestión y Mantenimiento Semanal", [
        ("Mantenimiento preventivo", [
            ("Test de mantecadora en vacío: verificar temperatura y ciclo", "Obrador"),
            ("Verificar niveles de refrigerante en equipos de frío", "Obrador"),
            ("Calibrar balanzas de obrador (pesos patrón)", "Obrador"),
            ("Revisar estado de utensilios: espátulas, porcionadores, moldes", "Obrador"),
        ]),
        ("Inventario y rotación", [
            ("Inventario rápido de materias primas clave", "Almacén"),
            ("Verificar rotación de sabores en vitrina (retirar poco demandados)", "Vitrina"),
            ("Comprobar stock de envases y material de servicio", "Almacén"),
        ]),
    ])
    create_task_sheet(wb, "Tareas Mensuales", ADMIN_COLOR, "Tareas Mensuales", [
        ("Calibración y mantenimiento", [
            ("Calibrar termómetros y sondas de temperatura", "Obrador"),
            ("Mantenimiento preventivo del pasteurizador (juntas, resistencias)", "Obrador"),
            ("Revisar juntas de puertas de cámaras frigoríficas", "Cámara"),
            ("Verificar estado de aislamiento de cámaras y vitrinas", "Cámara"),
        ]),
        ("Inventario y control", [
            ("Inventario completo valorado: MP, envases, producto terminado", "Almacén"),
            ("Calcular food cost mensual real vs objetivo (<28-32%)", "Admin"),
            ("Revisar mermas acumuladas del mes por sabor", "Admin"),
        ]),
        ("Documentación y calidad", [
            ("Planificar carta estacional del mes siguiente", "Admin"),
            ("Revisar y archivar documentación APPCC del mes", "Admin"),
            ("Test de alérgenos: verificar etiquetado y trazabilidad", "Admin"),
            ("Limpieza profunda general del local (suelos, paredes, techos)", "Limpieza"),
            ("Actualizar fichas técnicas si hay cambios de receta", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx"))
    print("  ✓ 05-tareas-semanales-mensuales.xlsx")


# ═══════════════════════════════════════════════════════════
# 06 — EVENTOS Y TEMPORADA
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Eventos y Temporada", [
        "Cómo usar esta plantilla:",
        "▸ Planificación estacional para heladería artesanal.",
        "▸ Temporada alta (primavera-verano): máxima producción y carta ampliada.",
        "▸ Temporada baja (otoño-invierno): carta reducida, productos complementarios.",
        "▸ Catering y eventos: logística y servicio externo.",
        "",
        "Personalización:",
        "▸ Ajusta fechas y sabores según tu zona geográfica.",
        "▸ Usa como base para la planificación trimestral.",
    ])
    create_task_sheet(wb, "Temporada Alta", OBRADOR_COLOR, "Temporada Alta — Primavera / Verano", [
        ("Preparación pre-temporada (marzo-abril)", [
            ("Ampliar carta a 18-24 sabores: incorporar frutales de temporada", "Obrador"),
            ("Reforzar equipo: contratar personal 2 semanas antes de temporada", "Equipo"),
            ("Aumentar stock de materias primas +40-50% respecto a invierno", "Almacén"),
            ("Verificar capacidad de producción (mantecadora, pasteurizador, cámaras)", "Obrador"),
            ("Planificar horario ampliado de apertura", "Admin"),
        ]),
        ("Operaciones temporada alta (mayo-septiembre)", [
            ("Producir diariamente según demanda (2-3 tandas de mantecación)", "Obrador"),
            ("Montar y gestionar servicio de terraza/exterior", "Terraza"),
            ("Reforzar mostrador en horas punta (2 dependientes mínimo)", "Servicio"),
            ("Campaña en RRSS: fotos de sabores, stories, reels de producción", "RRSS"),
            ("Preparar envases para llevar: tarrinas 0.5L, 1L, helado por kg", "Mostrador"),
            ("Control diario de stock: fruta fresca, lácteos, azúcares", "Almacén"),
        ]),
    ])
    create_task_sheet(wb, "Temporada Baja", CAMARA_COLOR, "Temporada Baja — Otoño / Invierno", [
        ("Ajuste de carta y operaciones (octubre-febrero)", [
            ("Reducir carta a 10-14 sabores: chocolate, turrón, canela, especias", "Obrador"),
            ("Incorporar productos complementarios: chocolate caliente, gofres, crêpes", "Obrador"),
            ("Ajustar horarios y turnos de personal a demanda reducida", "Equipo"),
            ("Reducir frecuencia de producción (1 tanda cada 2-3 días)", "Obrador"),
        ]),
        ("Mantenimiento y desarrollo", [
            ("Mantenimiento profundo de toda la maquinaria", "Obrador"),
            ("Testear y desarrollar nuevas recetas para la próxima temporada", "Obrador"),
            ("Formación del equipo: técnicas, atención al cliente, alérgenos", "Equipo"),
            ("Renovar decoración y cartelería del local si necesario", "General"),
            ("Revisar y renegociar contratos con proveedores", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Catering Eventos", MOSTRADOR_COLOR, "Catering y Eventos", [
        ("Preparación del evento", [
            ("Diseñar carta catering: formatos (bola, vasito, tarrina mini, tarta helada)", "Admin"),
            ("Calcular producción: 2-3 bolas/persona, mínimo 4 sabores", "Obrador"),
            ("Producir con 24-48h de antelación y almacenar a -18°C", "Obrador"),
            ("Preparar kit de servicio externo: cubetas, espátulas, guantes, servilletas", "Logística"),
        ]),
        ("Logística y transporte", [
            ("Verificar cadena de frío para transporte (cajas isotérmicas, hielo seco)", "Logística"),
            ("Cargar vehículo con producto y material de servicio", "Logística"),
            ("Verificar temperatura de producto a la llegada (<-12°C)", "Logística"),
        ]),
        ("Servicio y cierre", [
            ("Montar mesa de servicio en el evento según briefing del cliente", "Servicio"),
            ("Servir helado manteniendo presentación y rotación de sabores", "Servicio"),
            ("Recoger material, limpiar zona y transportar de vuelta", "Logística"),
            ("Emitir factura y enviar encuesta de satisfacción al cliente", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "06-eventos-temporada.xlsx"))
    print("  ✓ 06-eventos-temporada.xlsx")


# ═══════════════════════════════════════════════════════════
# 07 — PLANTILLA PERSONALIZABLE
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Plantilla Personalizable", [
        "Cómo usar esta plantilla:",
        "▸ Tres plantillas en blanco para crear tus propios checklists.",
        "▸ Por zona: organiza tareas según zona de la heladería.",
        "▸ Por turno: divide tareas según turno de mañana/tarde.",
        "▸ Por perfil: asigna tareas según rol del empleado.",
        "",
        "Personalización:",
        "▸ Completa las tareas según las necesidades de tu negocio.",
        "▸ Puedes duplicar hojas para crear más plantillas.",
        "▸ Imprime en A4 y plastifica para uso diario.",
    ])
    blank_headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
    create_blank_template_sheet(wb, "Por Zona", OBRADOR_COLOR, "Plantilla Personalizable — Por Zona", blank_headers)
    create_blank_template_sheet(wb, "Por Turno", VITRINA_COLOR, "Plantilla Personalizable — Por Turno", blank_headers)
    create_blank_template_sheet(wb, "Por Perfil", MOSTRADOR_COLOR, "Plantilla Personalizable — Por Perfil", blank_headers)
    wb.save(os.path.join(OUTPUT_DIR, "07-plantilla-personalizable.xlsx"))
    print("  ✓ 07-plantilla-personalizable.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS-01 — BRIEFING DIARIO DE HELADERÍA
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS-01 · Briefing Diario de Heladería", [
        "Cómo usar esta plantilla:",
        "▸ Reunión rápida de inicio de turno (5-10 minutos).",
        "▸ Producción: sabores a producir, mezclas listas, agotados.",
        "▸ Equipo: asignación de posiciones, eventos especiales.",
        "▸ Objetivos: ventas del día, sabor estrella, mejora continua.",
        "",
        "Personalización:",
        "▸ Imprime cada día y completa antes del briefing.",
        "▸ Cuelga en obrador y mostrador para referencia durante el turno.",
    ])
    create_task_sheet(wb, "Briefing Diario", ADMIN_COLOR, "Briefing Diario de Heladería", [
        ("Producción del día", [
            ("Revisar sabores a producir hoy (plan de producción)", "Obrador"),
            ("Verificar mezclas listas en maduración para manteccar", "Maduración"),
            ("Identificar sabores agotados o bajo mínimo en vitrina", "Vitrina"),
            ("Confirmar recepción de materias primas pendientes", "Almacén"),
            ("Comunicar si hay nueva receta o prueba de sabor", "Obrador"),
        ]),
        ("Equipo y servicio", [
            ("Confirmar personal presente y asignación de posiciones", "Equipo"),
            ("Informar de eventos, pedidos especiales o catering del día", "Admin"),
            ("Comunicar promociones, sabor nuevo o degustación del día", "Servicio"),
            ("Revisar incidencias del turno anterior que requieran acción", "Admin"),
        ]),
        ("Objetivos del turno", [
            ("Comunicar objetivo de ventas del día", "Admin"),
            ("Definir sabor estrella a recomendar a los clientes", "Servicio"),
            ("Asignar tarea de mejora del día (limpieza extra, reposición, etc.)", "General"),
            ("Recordar foco en atención al cliente y alérgenos", "Servicio"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))
    print("  ✓ BONUS-01-briefing-servicio.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS-02 — CALENDARIO ANUAL DE HELADERÍA
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS-02 · Calendario Anual de Heladería", [
        "Cómo usar esta plantilla:",
        "▸ Planificación anual con fechas clave y sabores temáticos.",
        "▸ Organiza producción estacional, campañas y eventos.",
        "▸ Marca ✓ las acciones completadas cada mes.",
        "",
        "Personalización:",
        "▸ Ajusta sabores y fechas según tu zona geográfica.",
        "▸ Añade eventos locales y fiestas de tu ciudad.",
        "▸ Planifica con 1 mes de antelación mínimo.",
    ])
    months_data = [
        ("Enero", "Reyes: roscón helado edición limitada. Inicio temporada baja, reducir carta.", "Turrón, Roscón, Canela, Chocolate", "Baja"),
        ("Febrero", "San Valentín: edición limitada (corazón, rosa, frutos rojos). Desarrollar recetas primavera.", "Frutos rojos, Rosa, Chocolate blanco", "Baja"),
        ("Marzo", "Semana Santa: sabores temáticos (torrija, mona). Preparar temporada alta: stock, equipo.", "Torrija, Mona, Vainilla", "Transición"),
        ("Abril", "Inicio primavera: ampliar carta con frutales. Reforzar equipo, formación.", "Fresa, Limón, Mango, Yogur", "Transición"),
        ("Mayo", "Inicio temporada alta. Carta completa 18-24 sabores. Abrir terraza.", "Fresa, Pistacho, Mango, Frambuesa", "Alta"),
        ("Junio", "Inicio verano: máxima producción. Día de la Madre: tartas heladas. Campaña RRSS.", "Melocotón, Cereza, Melón, Coco", "Alta"),
        ("Julio", "Máxima producción y ventas. Horario ampliado. Refuerzo personal.", "Sandía, Melocotón, Tropical, Sorbetes", "Alta"),
        ("Agosto", "Pico de temporada. Stock +50%. Catering eventos de verano.", "Higo, Sandía, Granizado, Fruta pasión", "Alta"),
        ("Septiembre", "Transición: reducir gradualmente carta. Vuelta al cole: promociones familia.", "Higo, Uva, Dulce de leche, Caramelo", "Transición"),
        ("Octubre", "Halloween: sabores temáticos (calabaza, mora). Reducir carta a 10-14.", "Calabaza, Mora, Castañas, Chocolate", "Baja"),
        ("Noviembre", "Temporada baja. Productos complementarios: chocolate caliente, gofres. Mantenimiento.", "Turrón, Chocolate, Especias, Avellana", "Baja"),
        ("Diciembre", "Navidad: turrón, polvorón, edición limitada. Catering empresas fin de año. Tartas heladas.", "Turrón, Polvorón, Mazapán, Canela", "Baja"),
    ]
    create_calendar_sheet(wb, "Calendario Anual", ADMIN_COLOR, "Calendario Anual de Heladería Artesanal", months_data)
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx"))
    print("  ✓ BONUS-02-calendario-anual-tareas.xlsx")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\n🍦 Generando Kit de Tareas — Heladería Artesanal\n   → {OUTPUT_DIR}\n")
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_bonus_01()
    gen_bonus_02()
    print(f"\n✅ 9 archivos generados en {OUTPUT_DIR}\n")
