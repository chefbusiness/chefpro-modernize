#!/usr/bin/env python3
"""
Kit de Escandallos Pro — Excel Generator
Generates all 11 professional templates + 1 bonus template.
Uses openpyxl with real formulas, formatting, and data validation.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

# ── Brand Colors ──────────────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
LIGHT_GOLD = "FFF8DC"
RED_ALERT = "FF4444"
GREEN_OK = "28A745"

# ── Styles ────────────────────────────────────────────────────
title_font = Font(name="Calibri", size=18, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888")
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
data_font = Font(name="Calibri", size=11, color="333333")
formula_font = Font(name="Calibri", size=11, color="333333", bold=True)
total_font = Font(name="Calibri", size=12, bold=True, color=DARK_BG)
total_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
light_row = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
alt_row = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-escandallos"
)

# ── Merma Data ────────────────────────────────────────────────
MERMAS = {
    "Carne roja": 0.20,
    "Aves": 0.25,
    "Pescado": 0.35,
    "Marisco": 0.45,
    "Verdura hoja": 0.25,
    "Verdura raíz": 0.12,
    "Fruta": 0.15,
    "Lácteos": 0.03,
    "Secos/granos": 0.02,
    "Congelados": 0.07,
    "Pan/bollería": 0.10,
    "Huevos": 0.11,
    "Aceites/grasas": 0.03,
    "Especias/hierbas": 0.20,
    "Chocolate/cacao": 0.05,
    "Bebidas/licores": 0.02,
}

UNIDADES = ["kg", "g", "L", "ml", "cl", "ud", "docena", "manojo", "sobre", "lata", "botella"]

CATEGORIAS_LIST = list(MERMAS.keys())

FOOD_COST_TARGETS = {
    "Fine Dining": {"fc_min": 0.25, "fc_max": 0.28, "mult": 3.8},
    "Casual Dining": {"fc_min": 0.28, "fc_max": 0.32, "mult": 3.33},
    "Fast Casual": {"fc_min": 0.30, "fc_max": 0.35, "mult": 3.0},
    "Cafetería": {"fc_min": 0.25, "fc_max": 0.30, "mult": 3.5},
    "Catering": {"fc_min": 0.30, "fc_max": 0.40, "mult": 2.85},
    "Food Truck": {"fc_min": 0.28, "fc_max": 0.35, "mult": 3.1},
    "Hotel F&B": {"fc_min": 0.30, "fc_max": 0.35, "mult": 3.1},
    "Pastelería": {"fc_min": 0.20, "fc_max": 0.30, "mult": 4.0},
    "Bar/Cocktails": {"fc_min": 0.18, "fc_max": 0.25, "mult": 4.5},
}


def add_instructions_sheet(wb, product_title, instructions_text):
    """Add a branded instructions sheet as the first tab."""
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80

    # Title
    ws.merge_cells("B2:B2")
    cell = ws["B2"]
    cell.value = f"📋 {product_title}"
    cell.font = title_font
    cell.alignment = left_align

    ws["B3"].value = "Kit de Escandallos Pro — AI Chef Pro"
    ws["B3"].font = subtitle_font

    ws["B4"].value = "www.aichef.pro"
    ws["B4"].font = Font(name="Calibri", size=10, color="888888", italic=True)

    # Instructions
    row = 6
    for line in instructions_text:
        ws[f"B{row}"].value = line
        if line.startswith("▸") or line.startswith("●"):
            ws[f"B{row}"].font = Font(name="Calibri", size=11, color="333333")
        elif line.startswith("—"):
            ws[f"B{row}"].font = Font(name="Calibri", size=11, color="888888", italic=True)
        elif line == "":
            pass
        else:
            ws[f"B{row}"].font = Font(name="Calibri", size=12, bold=True, color=DARK_BG)
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Footer
    row += 1
    ws[f"B{row}"].value = "© 2026 AI Chef Pro · Todos los derechos reservados"
    ws[f"B{row}"].font = Font(name="Calibri", size=9, color="AAAAAA")

    return ws


def setup_escandallo_sheet(wb, sheet_name, title, ingredients, tab_color="FFD700"):
    """Create a standard escandallo sheet with headers, data validation, and formulas."""
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # Column widths
    widths = {"A": 25, "B": 16, "C": 12, "D": 14, "E": 12, "F": 12, "G": 10, "H": 14, "I": 14, "J": 3, "K": 20, "L": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Image reference area (columns K-L, rows 4-14) ──
    img_border = Border(
        left=Side(style="medium", color=GOLD),
        right=Side(style="medium", color=GOLD),
        top=Side(style="medium", color=GOLD),
        bottom=Side(style="medium", color=GOLD),
    )
    img_inner_border = Border(
        left=Side(style="thin", color=MEDIUM_GRAY),
        right=Side(style="thin", color=MEDIUM_GRAY),
        top=Side(style="thin", color=MEDIUM_GRAY),
        bottom=Side(style="thin", color=MEDIUM_GRAY),
    )
    img_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

    # Header
    ws.merge_cells("K3:L3")
    ws["K3"].value = "📷 Foto del Plato"
    ws["K3"].font = Font(name="Calibri", size=12, bold=True, color=GOLD)
    ws["K3"].alignment = center_align
    ws["K3"].border = img_border
    ws["L3"].border = img_border

    # Image placeholder area (rows 4-14)
    for r in range(4, 15):
        for c in ["K", "L"]:
            cell = ws[f"{c}{r}"]
            cell.fill = img_fill
            # outer border on edges
            left = Side(style="medium", color=GOLD) if c == "K" else Side(style="thin", color=MEDIUM_GRAY)
            right = Side(style="medium", color=GOLD) if c == "L" else Side(style="thin", color=MEDIUM_GRAY)
            top = Side(style="medium", color=GOLD) if r == 4 else Side(style="thin", color=MEDIUM_GRAY)
            bottom = Side(style="medium", color=GOLD) if r == 14 else Side(style="thin", color=MEDIUM_GRAY)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Center instruction text
    ws.merge_cells("K8:L9")
    ws["K8"].value = "Arrastra aquí la foto\ndel plato terminado"
    ws["K8"].font = Font(name="Calibri", size=11, color="AAAAAA", italic=True)
    ws["K8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Footer note
    ws.merge_cells("K15:L15")
    ws["K15"].value = "Insertar → Imagen → Desde archivo"
    ws["K15"].font = Font(name="Calibri", size=9, color="BBBBBB", italic=True)
    ws["K15"].alignment = center_align

    # Title row
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="Calibri", size=16, bold=True, color=GOLD)
    cell.alignment = left_align

    ws.merge_cells("A2:I2")
    ws["A2"].value = "Introduce tus ingredientes y cantidades. Las celdas verdes son editables."
    ws["A2"].font = subtitle_font

    # Headers (row 4)
    headers = [
        "Ingrediente", "Categoría", "Ud. Compra", "Precio/Ud (€)",
        "Cantidad", "Ud. Uso", "Merma (%)", "Cant. Bruta", "Coste (€)"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data validation
    cat_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORIAS_LIST) + '"',
        allow_blank=True
    )
    cat_validation.error = "Selecciona una categoría válida"
    cat_validation.errorTitle = "Categoría inválida"
    ws.add_data_validation(cat_validation)

    unit_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(UNIDADES) + '"',
        allow_blank=True
    )
    ws.add_data_validation(unit_validation)

    # Data rows with ingredients
    data_start = 5
    for i, ing in enumerate(ingredients):
        row = data_start + i
        fill = light_row if i % 2 == 0 else alt_row

        # A: Ingrediente
        ws.cell(row=row, column=1, value=ing["name"]).font = data_font
        ws.cell(row=row, column=1).fill = input_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = left_align

        # B: Categoría
        ws.cell(row=row, column=2, value=ing["cat"]).font = data_font
        ws.cell(row=row, column=2).fill = input_fill
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = center_align
        cat_validation.add(ws.cell(row=row, column=2))

        # C: Unidad compra
        ws.cell(row=row, column=3, value=ing["unit"]).font = data_font
        ws.cell(row=row, column=3).fill = input_fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align
        unit_validation.add(ws.cell(row=row, column=3))

        # D: Precio/Ud
        ws.cell(row=row, column=4, value=ing["price"]).font = data_font
        ws.cell(row=row, column=4).fill = input_fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0.00 €'
        ws.cell(row=row, column=4).alignment = right_align

        # E: Cantidad
        ws.cell(row=row, column=5, value=ing["qty"]).font = data_font
        ws.cell(row=row, column=5).fill = input_fill
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '0.000'
        ws.cell(row=row, column=5).alignment = right_align

        # F: Unidad uso
        ws.cell(row=row, column=6, value=ing.get("use_unit", ing["unit"])).font = data_font
        ws.cell(row=row, column=6).fill = input_fill
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = center_align
        unit_validation.add(ws.cell(row=row, column=6))

        # G: Merma %
        ws.cell(row=row, column=7, value=ing["merma"]).font = data_font
        ws.cell(row=row, column=7).fill = input_fill
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).number_format = '0%'
        ws.cell(row=row, column=7).alignment = center_align

        # H: Cantidad bruta (formula) = Cantidad / (1 - Merma)
        formula_h = f"=E{row}/(1-G{row})"
        ws.cell(row=row, column=8, value=formula_h).font = formula_font
        ws.cell(row=row, column=8).border = thin_border
        ws.cell(row=row, column=8).number_format = '0.000'
        ws.cell(row=row, column=8).alignment = right_align

        # I: Coste (formula) = Cantidad bruta × Precio/Ud
        formula_i = f"=H{row}*D{row}"
        ws.cell(row=row, column=9, value=formula_i).font = formula_font
        ws.cell(row=row, column=9).border = thin_border
        ws.cell(row=row, column=9).number_format = '#,##0.00 €'
        ws.cell(row=row, column=9).alignment = right_align

    # Add empty rows for user to fill (5 extra)
    last_data = data_start + len(ingredients)
    for i in range(5):
        row = last_data + i
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill if col <= 7 else PatternFill()
            if col == 2:
                cat_validation.add(cell)
            if col in (3, 6):
                unit_validation.add(cell)
            if col == 7:
                cell.number_format = '0%'
            if col == 8:
                cell.value = f"=IF(E{row}=\"\",\"\",E{row}/(1-G{row}))"
                cell.number_format = '0.000'
            if col == 9:
                cell.value = f"=IF(H{row}=\"\",\"\",H{row}*D{row})"
                cell.number_format = '#,##0.00 €'

    # ── Totals section ──
    totals_start = last_data + 5 + 1
    last_cost_row = last_data + 4  # last possible data row

    # Coste total ingredientes
    ws.merge_cells(f"A{totals_start}:H{totals_start}")
    ws.cell(row=totals_start, column=1, value="COSTE TOTAL INGREDIENTES").font = total_font
    ws.cell(row=totals_start, column=1).fill = total_fill
    ws.cell(row=totals_start, column=1).border = thin_border
    ws.cell(row=totals_start, column=1).alignment = right_align
    ws.cell(row=totals_start, column=9, value=f"=SUM(I{data_start}:I{last_cost_row})").font = total_font
    ws.cell(row=totals_start, column=9).fill = total_fill
    ws.cell(row=totals_start, column=9).border = thin_border
    ws.cell(row=totals_start, column=9).number_format = '#,##0.00 €'

    # Coste elaboración (10%)
    r = totals_start + 1
    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(row=r, column=1, value="Coste elaboración (%)").font = data_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=8, value=0.10).font = data_font
    ws.cell(row=r, column=8).fill = input_fill
    ws.cell(row=r, column=8).border = thin_border
    ws.cell(row=r, column=8).number_format = '0%'
    ws.cell(row=r, column=8).alignment = center_align
    ws.cell(row=r, column=9, value=f"=I{totals_start}*H{r}").font = formula_font
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=9).number_format = '#,##0.00 €'

    # Coste total plato
    r = totals_start + 2
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1, value="COSTE TOTAL DEL PLATO").font = Font(name="Calibri", size=13, bold=True, color=DARK_BG)
    ws.cell(row=r, column=1).fill = gold_fill
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=9, value=f"=I{totals_start}+I{totals_start+1}").font = Font(name="Calibri", size=13, bold=True)
    ws.cell(row=r, column=9).fill = gold_fill
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=9).number_format = '#,##0.00 €'

    # Food cost objetivo
    r = totals_start + 4
    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(row=r, column=1, value="Food Cost objetivo (%)").font = Font(name="Calibri", size=11, bold=True, color="333333")
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=8, value=0.30).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=r, column=8).fill = input_fill
    ws.cell(row=r, column=8).border = thin_border
    ws.cell(row=r, column=8).number_format = '0%'
    ws.cell(row=r, column=8).alignment = center_align

    # PVP sugerido sin IVA
    r = totals_start + 5
    coste_total_cell = f"I{totals_start+2}"
    fc_cell = f"H{totals_start+4}"
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1, value="PVP SUGERIDO (sin IVA)").font = total_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=9, value=f"={coste_total_cell}/{fc_cell}").font = total_font
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=9).number_format = '#,##0.00 €'

    # PVP con IVA (10%)
    r = totals_start + 6
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1, value="PVP CON IVA (10% hostelería)").font = total_font
    ws.cell(row=r, column=1).fill = gold_fill
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=9, value=f"=I{totals_start+5}*1.10").font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=r, column=9).fill = gold_fill
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=9).number_format = '#,##0.00 €'

    # Margen bruto €
    r = totals_start + 7
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1, value="Margen bruto (€)").font = data_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=9, value=f"=I{totals_start+5}-{coste_total_cell}").font = formula_font
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=9).number_format = '#,##0.00 €'

    # Margen bruto %
    r = totals_start + 8
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1, value="Margen bruto (%)").font = data_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = right_align
    ws.cell(row=r, column=9, value=f"=I{totals_start+7}/I{totals_start+5}").font = formula_font
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=9).number_format = '0.0%'

    # Freeze panes
    ws.freeze_panes = "A5"

    # Print setup
    ws.print_area = f"A1:I{totals_start + 8}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return ws, totals_start


# ══════════════════════════════════════════════════════════════
# TEMPLATE 1: Escandallo Estándar (Plato a la Carta)
# ══════════════════════════════════════════════════════════════
def create_escandallo_estandar():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo Estándar — Plato a la Carta", [
        "Cómo usar esta plantilla",
        "",
        "▸ Ve a la pestaña 'Escandallo' para empezar",
        "▸ Las celdas verdes son editables — introduce tus ingredientes y cantidades",
        "▸ Las fórmulas se calculan automáticamente (coste, merma, PVP)",
        "▸ La merma viene precargada por categoría de ingrediente (puedes ajustarla)",
        "▸ El Food Cost objetivo por defecto es 30% — cámbialo según tu tipo de establecimiento",
        "",
        "Columnas explicadas",
        "",
        "▸ Ingrediente: nombre del producto",
        "▸ Categoría: tipo de ingrediente (selecciona del desplegable)",
        "▸ Ud. Compra: unidad en la que compras (kg, L, ud...)",
        "▸ Precio/Ud: precio que pagas por esa unidad de compra",
        "▸ Cantidad: cantidad que necesitas para el plato (en la unidad de uso)",
        "▸ Ud. Uso: unidad en la que mides para el plato",
        "▸ Merma (%): porcentaje de desperdicio (precargado, editable)",
        "▸ Cant. Bruta: cantidad a comprar contando la merma (automático)",
        "▸ Coste (€): coste real del ingrediente con merma incluida (automático)",
        "",
        "Referencia de mermas estándar",
        "",
        "▸ Carne roja: 20% | Aves: 25% | Pescado: 35% | Marisco: 45%",
        "▸ Verdura hoja: 25% | Verdura raíz: 12% | Fruta: 15%",
        "▸ Lácteos: 3% | Secos: 2% | Congelados: 7%",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    ingredients = [
        {"name": "Solomillo de ternera", "cat": "Carne roja", "unit": "kg", "price": 28.50, "qty": 0.200, "merma": 0.20},
        {"name": "Patata agria", "cat": "Verdura raíz", "unit": "kg", "price": 1.80, "qty": 0.150, "merma": 0.12},
        {"name": "Espárrago verde", "cat": "Verdura hoja", "unit": "kg", "price": 6.50, "qty": 0.080, "merma": 0.25},
        {"name": "Mantequilla", "cat": "Lácteos", "unit": "kg", "price": 9.00, "qty": 0.030, "merma": 0.03},
        {"name": "Vino Pedro Ximénez", "cat": "Bebidas/licores", "unit": "L", "price": 8.50, "qty": 0.050, "merma": 0.02},
        {"name": "Fondo oscuro de ternera", "cat": "Congelados", "unit": "L", "price": 12.00, "qty": 0.080, "merma": 0.07},
        {"name": "Aceite de oliva virgen extra", "cat": "Aceites/grasas", "unit": "L", "price": 7.50, "qty": 0.030, "merma": 0.03},
        {"name": "Flor de sal", "cat": "Secos/granos", "unit": "kg", "price": 15.00, "qty": 0.003, "merma": 0.02},
        {"name": "Pimienta negra", "cat": "Especias/hierbas", "unit": "kg", "price": 25.00, "qty": 0.002, "merma": 0.20},
        {"name": "Microgreens decoración", "cat": "Verdura hoja", "unit": "ud", "price": 3.50, "qty": 1.000, "use_unit": "ud", "merma": 0.10},
    ]

    setup_escandallo_sheet(wb, "Escandallo", "Escandallo — Solomillo con PX y Espárragos", ingredients)

    path = os.path.join(OUTPUT_DIR, "01-escandallo-estandar.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 2: Menú Degustación
# ══════════════════════════════════════════════════════════════
def create_menu_degustacion():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Menú Degustación", [
        "Cómo usar esta plantilla",
        "",
        "▸ Cada pestaña es un pase del menú degustación",
        "▸ La pestaña 'Resumen' suma automáticamente todos los pases",
        "▸ Introduce ingredientes y cantidades en cada pestaña de pase",
        "▸ El PVP se calcula sobre el coste total de todos los pases",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    pases = [
        ("1. Aperitivo", "Aperitivo — Tartar de Atún con Aguacate", [
            {"name": "Atún rojo (lomo)", "cat": "Pescado", "unit": "kg", "price": 42.00, "qty": 0.080, "merma": 0.35},
            {"name": "Aguacate Hass", "cat": "Fruta", "unit": "ud", "price": 1.50, "qty": 0.500, "use_unit": "ud", "merma": 0.30},
            {"name": "Sésamo tostado", "cat": "Secos/granos", "unit": "kg", "price": 12.00, "qty": 0.005, "merma": 0.02},
            {"name": "Salsa de soja", "cat": "Bebidas/licores", "unit": "L", "price": 6.00, "qty": 0.015, "merma": 0.02},
            {"name": "Aceite de sésamo", "cat": "Aceites/grasas", "unit": "L", "price": 14.00, "qty": 0.010, "merma": 0.03},
        ]),
        ("2. Entrante", "Entrante — Vieira con Coliflor y Trufa", [
            {"name": "Vieira fresca (grande)", "cat": "Marisco", "unit": "kg", "price": 48.00, "qty": 0.060, "merma": 0.45},
            {"name": "Coliflor", "cat": "Verdura raíz", "unit": "kg", "price": 2.50, "qty": 0.100, "merma": 0.30},
            {"name": "Trufa negra", "cat": "Verdura raíz", "unit": "kg", "price": 800.00, "qty": 0.005, "merma": 0.10},
            {"name": "Nata 35% MG", "cat": "Lácteos", "unit": "L", "price": 3.80, "qty": 0.050, "merma": 0.03},
            {"name": "Mantequilla", "cat": "Lácteos", "unit": "kg", "price": 9.00, "qty": 0.020, "merma": 0.03},
        ]),
        ("3. Pescado", "Pescado — Lubina con Beurre Blanc", [
            {"name": "Lubina salvaje", "cat": "Pescado", "unit": "kg", "price": 32.00, "qty": 0.160, "merma": 0.40},
            {"name": "Chalota", "cat": "Verdura raíz", "unit": "kg", "price": 4.50, "qty": 0.030, "merma": 0.15},
            {"name": "Vino blanco seco", "cat": "Bebidas/licores", "unit": "L", "price": 5.00, "qty": 0.040, "merma": 0.02},
            {"name": "Mantequilla", "cat": "Lácteos", "unit": "kg", "price": 9.00, "qty": 0.040, "merma": 0.03},
            {"name": "Hinojo fresco", "cat": "Verdura hoja", "unit": "kg", "price": 5.00, "qty": 0.060, "merma": 0.25},
        ]),
        ("4. Carne", "Carne — Carrillera Ibérica al Vino Tinto", [
            {"name": "Carrillera ibérica", "cat": "Carne roja", "unit": "kg", "price": 18.00, "qty": 0.200, "merma": 0.15},
            {"name": "Vino tinto reserva", "cat": "Bebidas/licores", "unit": "L", "price": 6.00, "qty": 0.100, "merma": 0.02},
            {"name": "Zanahoria", "cat": "Verdura raíz", "unit": "kg", "price": 1.20, "qty": 0.050, "merma": 0.12},
            {"name": "Puré de boniato", "cat": "Verdura raíz", "unit": "kg", "price": 2.50, "qty": 0.100, "merma": 0.15},
            {"name": "Romero fresco", "cat": "Especias/hierbas", "unit": "manojo", "price": 1.00, "qty": 0.500, "use_unit": "ud", "merma": 0.20},
        ]),
        ("5. Postre", "Postre — Esfera de Chocolate con Fruta de la Pasión", [
            {"name": "Chocolate negro 70%", "cat": "Chocolate/cacao", "unit": "kg", "price": 18.00, "qty": 0.060, "merma": 0.05},
            {"name": "Fruta de la pasión", "cat": "Fruta", "unit": "kg", "price": 14.00, "qty": 0.040, "merma": 0.50},
            {"name": "Nata 35% MG", "cat": "Lácteos", "unit": "L", "price": 3.80, "qty": 0.060, "merma": 0.03},
            {"name": "Azúcar", "cat": "Secos/granos", "unit": "kg", "price": 1.20, "qty": 0.025, "merma": 0.02},
            {"name": "Manteca de cacao", "cat": "Chocolate/cacao", "unit": "kg", "price": 22.00, "qty": 0.015, "merma": 0.05},
        ]),
    ]

    totals_cells = []
    for sheet_name, title, ingredients in pases:
        ws, totals_start = setup_escandallo_sheet(wb, sheet_name, title, ingredients)
        totals_cells.append((sheet_name, totals_start + 2))  # coste total del plato row

    # ── Summary sheet ──
    ws_sum = wb.create_sheet(title="Resumen")
    ws_sum.sheet_properties.tabColor = "FF4444"

    ws_sum.column_dimensions["A"].width = 5
    ws_sum.column_dimensions["B"].width = 40
    ws_sum.column_dimensions["C"].width = 18

    ws_sum.merge_cells("B2:C2")
    ws_sum["B2"].value = "Resumen — Menú Degustación"
    ws_sum["B2"].font = title_font

    ws_sum["B4"].value = "Pase"
    ws_sum["B4"].font = header_font
    ws_sum["B4"].fill = header_fill
    ws_sum["B4"].border = thin_border
    ws_sum["C4"].value = "Coste (€)"
    ws_sum["C4"].font = header_font
    ws_sum["C4"].fill = header_fill
    ws_sum["C4"].border = thin_border
    ws_sum["C4"].alignment = right_align

    for i, (sheet_name, total_row) in enumerate(totals_cells):
        row = 5 + i
        ws_sum.cell(row=row, column=2, value=sheet_name).font = data_font
        ws_sum.cell(row=row, column=2).border = thin_border
        ws_sum.cell(row=row, column=3, value=f"='{sheet_name}'!I{total_row}").font = formula_font
        ws_sum.cell(row=row, column=3).border = thin_border
        ws_sum.cell(row=row, column=3).number_format = '#,##0.00 €'
        ws_sum.cell(row=row, column=3).alignment = right_align

    total_row_sum = 5 + len(pases)
    ws_sum.cell(row=total_row_sum, column=2, value="COSTE TOTAL MENÚ").font = total_font
    ws_sum.cell(row=total_row_sum, column=2).fill = gold_fill
    ws_sum.cell(row=total_row_sum, column=2).border = thin_border
    ws_sum.cell(row=total_row_sum, column=3, value=f"=SUM(C5:C{total_row_sum-1})").font = Font(name="Calibri", size=14, bold=True)
    ws_sum.cell(row=total_row_sum, column=3).fill = gold_fill
    ws_sum.cell(row=total_row_sum, column=3).border = thin_border
    ws_sum.cell(row=total_row_sum, column=3).number_format = '#,##0.00 €'

    # PVP section
    r = total_row_sum + 2
    ws_sum.cell(row=r, column=2, value="Food Cost objetivo (%)").font = Font(name="Calibri", size=11, bold=True)
    ws_sum.cell(row=r, column=2).border = thin_border
    ws_sum.cell(row=r, column=3, value=0.25).font = data_font
    ws_sum.cell(row=r, column=3).fill = input_fill
    ws_sum.cell(row=r, column=3).border = thin_border
    ws_sum.cell(row=r, column=3).number_format = '0%'

    r += 1
    ws_sum.cell(row=r, column=2, value="PVP SUGERIDO (sin IVA)").font = total_font
    ws_sum.cell(row=r, column=2).border = thin_border
    ws_sum.cell(row=r, column=3, value=f"=C{total_row_sum}/C{r-1}").font = total_font
    ws_sum.cell(row=r, column=3).border = thin_border
    ws_sum.cell(row=r, column=3).number_format = '#,##0.00 €'

    r += 1
    ws_sum.cell(row=r, column=2, value="PVP CON IVA (10%)").font = Font(name="Calibri", size=14, bold=True)
    ws_sum.cell(row=r, column=2).fill = gold_fill
    ws_sum.cell(row=r, column=2).border = thin_border
    ws_sum.cell(row=r, column=3, value=f"=C{r-1}*1.10").font = Font(name="Calibri", size=14, bold=True)
    ws_sum.cell(row=r, column=3).fill = gold_fill
    ws_sum.cell(row=r, column=3).border = thin_border
    ws_sum.cell(row=r, column=3).number_format = '#,##0.00 €'

    path = os.path.join(OUTPUT_DIR, "02-menu-degustacion.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 3: Menú del Día
# ══════════════════════════════════════════════════════════════
def create_menu_del_dia():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Menú del Día", [
        "Cómo usar esta plantilla",
        "",
        "▸ Pestaña 'Primer Plato': escandallo del primer plato",
        "▸ Pestaña 'Segundo Plato': escandallo del segundo plato",
        "▸ Pestaña 'Postre': escandallo del postre",
        "▸ Pestaña 'Resumen': coste total del menú + PVP sugerido",
        "▸ Incluye pan, bebida y café como costes fijos",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    platos = [
        ("Primer Plato", "Primer Plato — Ensalada Mediterránea", [
            {"name": "Lechuga variada", "cat": "Verdura hoja", "unit": "kg", "price": 3.50, "qty": 0.080, "merma": 0.25},
            {"name": "Tomate pera", "cat": "Verdura hoja", "unit": "kg", "price": 2.80, "qty": 0.100, "merma": 0.10},
            {"name": "Pepino", "cat": "Verdura raíz", "unit": "kg", "price": 1.50, "qty": 0.060, "merma": 0.12},
            {"name": "Cebolla morada", "cat": "Verdura raíz", "unit": "kg", "price": 1.80, "qty": 0.030, "merma": 0.12},
            {"name": "Aceitunas negras", "cat": "Congelados", "unit": "kg", "price": 5.00, "qty": 0.020, "merma": 0.02},
            {"name": "Queso feta", "cat": "Lácteos", "unit": "kg", "price": 12.00, "qty": 0.040, "merma": 0.03},
            {"name": "AOVE", "cat": "Aceites/grasas", "unit": "L", "price": 7.50, "qty": 0.020, "merma": 0.03},
        ]),
        ("Segundo Plato", "Segundo Plato — Pollo al Horno con Patatas", [
            {"name": "Contramuslo de pollo", "cat": "Aves", "unit": "kg", "price": 5.50, "qty": 0.250, "merma": 0.25},
            {"name": "Patata", "cat": "Verdura raíz", "unit": "kg", "price": 1.20, "qty": 0.200, "merma": 0.12},
            {"name": "Pimiento rojo", "cat": "Verdura hoja", "unit": "kg", "price": 3.00, "qty": 0.060, "merma": 0.15},
            {"name": "Ajo", "cat": "Especias/hierbas", "unit": "kg", "price": 6.00, "qty": 0.010, "merma": 0.15},
            {"name": "Romero", "cat": "Especias/hierbas", "unit": "manojo", "price": 1.00, "qty": 0.250, "use_unit": "ud", "merma": 0.20},
            {"name": "AOVE", "cat": "Aceites/grasas", "unit": "L", "price": 7.50, "qty": 0.025, "merma": 0.03},
        ]),
        ("Postre", "Postre — Flan Casero", [
            {"name": "Huevos", "cat": "Huevos", "unit": "docena", "price": 3.60, "qty": 0.250, "use_unit": "ud", "merma": 0.11},
            {"name": "Leche entera", "cat": "Lácteos", "unit": "L", "price": 1.10, "qty": 0.150, "merma": 0.03},
            {"name": "Azúcar", "cat": "Secos/granos", "unit": "kg", "price": 1.20, "qty": 0.040, "merma": 0.02},
            {"name": "Vainilla (extracto)", "cat": "Especias/hierbas", "unit": "L", "price": 35.00, "qty": 0.003, "merma": 0.05},
            {"name": "Caramelo líquido", "cat": "Secos/granos", "unit": "kg", "price": 4.00, "qty": 0.015, "merma": 0.05},
        ]),
    ]

    totals_cells = []
    for sheet_name, title, ingredients in platos:
        ws, totals_start = setup_escandallo_sheet(wb, sheet_name, title, ingredients)
        totals_cells.append((sheet_name, totals_start + 2))

    # Summary
    ws_sum = wb.create_sheet(title="Resumen Menú")
    ws_sum.sheet_properties.tabColor = "FF4444"
    ws_sum.column_dimensions["B"].width = 40
    ws_sum.column_dimensions["C"].width = 18

    ws_sum["B2"].value = "Resumen — Menú del Día"
    ws_sum["B2"].font = title_font

    ws_sum["B4"].value = "Concepto"
    ws_sum["B4"].font = header_font
    ws_sum["B4"].fill = header_fill
    ws_sum["B4"].border = thin_border
    ws_sum["C4"].value = "Coste (€)"
    ws_sum["C4"].font = header_font
    ws_sum["C4"].fill = header_fill
    ws_sum["C4"].border = thin_border

    row = 5
    for sheet_name, total_row in totals_cells:
        ws_sum.cell(row=row, column=2, value=sheet_name).font = data_font
        ws_sum.cell(row=row, column=2).border = thin_border
        ws_sum.cell(row=row, column=3, value=f"='{sheet_name}'!I{total_row}").font = formula_font
        ws_sum.cell(row=row, column=3).border = thin_border
        ws_sum.cell(row=row, column=3).number_format = '#,##0.00 €'
        row += 1

    # Extras fijos
    extras = [("Pan (por comensal)", 0.15), ("Bebida (agua/vino copa)", 0.40), ("Café", 0.20)]
    for name, cost in extras:
        ws_sum.cell(row=row, column=2, value=name).font = data_font
        ws_sum.cell(row=row, column=2).border = thin_border
        ws_sum.cell(row=row, column=3, value=cost).font = data_font
        ws_sum.cell(row=row, column=3).fill = input_fill
        ws_sum.cell(row=row, column=3).border = thin_border
        ws_sum.cell(row=row, column=3).number_format = '#,##0.00 €'
        row += 1

    # Total
    ws_sum.cell(row=row, column=2, value="COSTE TOTAL MENÚ DEL DÍA").font = total_font
    ws_sum.cell(row=row, column=2).fill = gold_fill
    ws_sum.cell(row=row, column=2).border = thin_border
    ws_sum.cell(row=row, column=3, value=f"=SUM(C5:C{row-1})").font = Font(name="Calibri", size=14, bold=True)
    ws_sum.cell(row=row, column=3).fill = gold_fill
    ws_sum.cell(row=row, column=3).border = thin_border
    ws_sum.cell(row=row, column=3).number_format = '#,##0.00 €'

    total_row_ref = row
    row += 2
    ws_sum.cell(row=row, column=2, value="Food Cost objetivo (%)").font = Font(name="Calibri", size=11, bold=True)
    ws_sum.cell(row=row, column=2).border = thin_border
    ws_sum.cell(row=row, column=3, value=0.33).fill = input_fill
    ws_sum.cell(row=row, column=3).border = thin_border
    ws_sum.cell(row=row, column=3).number_format = '0%'

    row += 1
    ws_sum.cell(row=row, column=2, value="PVP SUGERIDO MENÚ (sin IVA)").font = total_font
    ws_sum.cell(row=row, column=2).border = thin_border
    ws_sum.cell(row=row, column=3, value=f"=C{total_row_ref}/C{row-1}").font = total_font
    ws_sum.cell(row=row, column=3).border = thin_border
    ws_sum.cell(row=row, column=3).number_format = '#,##0.00 €'

    row += 1
    ws_sum.cell(row=row, column=2, value="PVP CON IVA (10%)").font = Font(name="Calibri", size=14, bold=True)
    ws_sum.cell(row=row, column=2).fill = gold_fill
    ws_sum.cell(row=row, column=2).border = thin_border
    ws_sum.cell(row=row, column=3, value=f"=C{row-1}*1.10").font = Font(name="Calibri", size=14, bold=True)
    ws_sum.cell(row=row, column=3).fill = gold_fill
    ws_sum.cell(row=row, column=3).border = thin_border
    ws_sum.cell(row=row, column=3).number_format = '#,##0.00 €'

    path = os.path.join(OUTPUT_DIR, "03-menu-del-dia.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 4: Cocktails y Bebidas
# ══════════════════════════════════════════════════════════════
def create_cocktails():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Cocktails y Bebidas", [
        "Cómo usar esta plantilla",
        "",
        "▸ Cada pestaña es un cocktail diferente",
        "▸ Las cantidades están en cl/ml (medidas de bar)",
        "▸ La merma en bebidas es mínima (2-5%)",
        "▸ El food cost objetivo en bar suele ser 18-25%",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    cocktails = [
        ("Gin Tonic Premium", "Gin Tonic Premium", [
            {"name": "Ginebra premium", "cat": "Bebidas/licores", "unit": "L", "price": 28.00, "qty": 0.050, "use_unit": "cl", "merma": 0.02},
            {"name": "Tónica premium", "cat": "Bebidas/licores", "unit": "ud", "price": 1.50, "qty": 1.000, "use_unit": "ud", "merma": 0.00},
            {"name": "Limón", "cat": "Fruta", "unit": "kg", "price": 2.50, "qty": 0.020, "merma": 0.30},
            {"name": "Pepino", "cat": "Verdura raíz", "unit": "kg", "price": 1.50, "qty": 0.010, "merma": 0.12},
            {"name": "Enebro (bayas)", "cat": "Especias/hierbas", "unit": "kg", "price": 30.00, "qty": 0.002, "merma": 0.05},
            {"name": "Hielo", "cat": "Congelados", "unit": "kg", "price": 0.50, "qty": 0.150, "merma": 0.00},
        ]),
        ("Mojito Clásico", "Mojito Clásico", [
            {"name": "Ron blanco", "cat": "Bebidas/licores", "unit": "L", "price": 15.00, "qty": 0.060, "use_unit": "cl", "merma": 0.02},
            {"name": "Lima", "cat": "Fruta", "unit": "kg", "price": 4.50, "qty": 0.030, "merma": 0.30},
            {"name": "Hierbabuena", "cat": "Especias/hierbas", "unit": "manojo", "price": 1.20, "qty": 0.500, "use_unit": "ud", "merma": 0.30},
            {"name": "Azúcar", "cat": "Secos/granos", "unit": "kg", "price": 1.20, "qty": 0.015, "merma": 0.02},
            {"name": "Soda/agua con gas", "cat": "Bebidas/licores", "unit": "L", "price": 0.80, "qty": 0.060, "merma": 0.00},
            {"name": "Hielo", "cat": "Congelados", "unit": "kg", "price": 0.50, "qty": 0.150, "merma": 0.00},
        ]),
        ("Margarita", "Margarita Clásica", [
            {"name": "Tequila reposado", "cat": "Bebidas/licores", "unit": "L", "price": 22.00, "qty": 0.050, "use_unit": "cl", "merma": 0.02},
            {"name": "Triple Sec / Cointreau", "cat": "Bebidas/licores", "unit": "L", "price": 18.00, "qty": 0.025, "use_unit": "cl", "merma": 0.02},
            {"name": "Lima (zumo)", "cat": "Fruta", "unit": "kg", "price": 4.50, "qty": 0.030, "merma": 0.50},
            {"name": "Sal (para borde)", "cat": "Secos/granos", "unit": "kg", "price": 0.80, "qty": 0.005, "merma": 0.50},
            {"name": "Hielo", "cat": "Congelados", "unit": "kg", "price": 0.50, "qty": 0.100, "merma": 0.00},
        ]),
        ("Aperol Spritz", "Aperol Spritz", [
            {"name": "Aperol", "cat": "Bebidas/licores", "unit": "L", "price": 14.00, "qty": 0.060, "use_unit": "cl", "merma": 0.02},
            {"name": "Prosecco", "cat": "Bebidas/licores", "unit": "L", "price": 8.00, "qty": 0.090, "use_unit": "cl", "merma": 0.05},
            {"name": "Soda", "cat": "Bebidas/licores", "unit": "L", "price": 0.80, "qty": 0.030, "use_unit": "cl", "merma": 0.00},
            {"name": "Naranja (rodaja)", "cat": "Fruta", "unit": "kg", "price": 2.00, "qty": 0.020, "merma": 0.25},
            {"name": "Hielo", "cat": "Congelados", "unit": "kg", "price": 0.50, "qty": 0.150, "merma": 0.00},
        ]),
    ]

    for sheet_name, title, ingredients in cocktails:
        setup_escandallo_sheet(wb, sheet_name, title, ingredients)
        # Override food cost to 20% for bar
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 0.30 and isinstance(cell.value, float):
                    # Find the food cost cell and set to 20%
                    cell.value = 0.20
                    break

    path = os.path.join(OUTPUT_DIR, "04-cocktails-bebidas.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 5: Pastelería
# ══════════════════════════════════════════════════════════════
def create_pasteleria():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Pastelería y Repostería", [
        "Cómo usar esta plantilla",
        "",
        "▸ Incluye campo de 'Rendimiento' (unidades por receta)",
        "▸ El coste se calcula por unidad producida",
        "▸ Mermas de pastelería son generalmente más bajas",
        "▸ Food cost objetivo en pastelería: 20-30%",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    recetas = [
        ("Tarta Chocolate", "Tarta de Chocolate 70% — 12 raciones", [
            {"name": "Chocolate negro 70%", "cat": "Chocolate/cacao", "unit": "kg", "price": 18.00, "qty": 0.400, "merma": 0.05},
            {"name": "Mantequilla", "cat": "Lácteos", "unit": "kg", "price": 9.00, "qty": 0.250, "merma": 0.03},
            {"name": "Huevos", "cat": "Huevos", "unit": "docena", "price": 3.60, "qty": 0.500, "use_unit": "ud", "merma": 0.11},
            {"name": "Azúcar", "cat": "Secos/granos", "unit": "kg", "price": 1.20, "qty": 0.200, "merma": 0.02},
            {"name": "Harina floja", "cat": "Secos/granos", "unit": "kg", "price": 0.90, "qty": 0.120, "merma": 0.02},
            {"name": "Nata 35% MG", "cat": "Lácteos", "unit": "L", "price": 3.80, "qty": 0.200, "merma": 0.03},
            {"name": "Cacao en polvo", "cat": "Chocolate/cacao", "unit": "kg", "price": 12.00, "qty": 0.030, "merma": 0.05},
            {"name": "Sal", "cat": "Secos/granos", "unit": "kg", "price": 0.60, "qty": 0.003, "merma": 0.02},
        ]),
        ("Croissants", "Croissants de Mantequilla — 20 unidades", [
            {"name": "Harina de fuerza", "cat": "Secos/granos", "unit": "kg", "price": 1.10, "qty": 0.500, "merma": 0.02},
            {"name": "Mantequilla seca (82% MG)", "cat": "Lácteos", "unit": "kg", "price": 12.00, "qty": 0.280, "merma": 0.03},
            {"name": "Leche entera", "cat": "Lácteos", "unit": "L", "price": 1.10, "qty": 0.150, "merma": 0.03},
            {"name": "Azúcar", "cat": "Secos/granos", "unit": "kg", "price": 1.20, "qty": 0.060, "merma": 0.02},
            {"name": "Levadura fresca", "cat": "Lácteos", "unit": "kg", "price": 5.00, "qty": 0.020, "merma": 0.05},
            {"name": "Sal", "cat": "Secos/granos", "unit": "kg", "price": 0.60, "qty": 0.010, "merma": 0.02},
            {"name": "Huevo (para pintar)", "cat": "Huevos", "unit": "docena", "price": 3.60, "qty": 0.083, "use_unit": "ud", "merma": 0.11},
        ]),
        ("Macarons", "Macarons de Frambuesa — 30 unidades", [
            {"name": "Harina de almendra", "cat": "Secos/granos", "unit": "kg", "price": 14.00, "qty": 0.150, "merma": 0.02},
            {"name": "Azúcar glas", "cat": "Secos/granos", "unit": "kg", "price": 2.50, "qty": 0.150, "merma": 0.02},
            {"name": "Claras de huevo", "cat": "Huevos", "unit": "kg", "price": 6.00, "qty": 0.110, "merma": 0.05},
            {"name": "Azúcar", "cat": "Secos/granos", "unit": "kg", "price": 1.20, "qty": 0.150, "merma": 0.02},
            {"name": "Frambuesa congelada", "cat": "Congelados", "unit": "kg", "price": 9.00, "qty": 0.100, "merma": 0.07},
            {"name": "Chocolate blanco", "cat": "Chocolate/cacao", "unit": "kg", "price": 15.00, "qty": 0.080, "merma": 0.05},
            {"name": "Colorante rojo", "cat": "Especias/hierbas", "unit": "ud", "price": 4.00, "qty": 0.100, "use_unit": "ud", "merma": 0.00},
        ]),
    ]

    for sheet_name, title, ingredients in recetas:
        ws, _ = setup_escandallo_sheet(wb, sheet_name, title, ingredients)
        # Override food cost to 25% for pastry
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 0.30 and isinstance(cell.value, float):
                    cell.value = 0.25
                    break

    path = os.path.join(OUTPUT_DIR, "05-pasteleria.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 6: Catering por Persona
# ══════════════════════════════════════════════════════════════
def create_catering():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Catering por Persona", [
        "Cómo usar esta plantilla",
        "",
        "▸ Cada pestaña calcula el coste POR PERSONA",
        "▸ La pestaña 'Presupuesto' multiplica por número de comensales",
        "▸ Incluye servicio, personal y logística como costes fijos",
        "▸ Food cost objetivo en catering: 30-40%",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    # Cocktail catering per person
    cocktail_items = [
        {"name": "Jamón ibérico (loncheado)", "cat": "Carne roja", "unit": "kg", "price": 65.00, "qty": 0.040, "merma": 0.05},
        {"name": "Queso manchego curado", "cat": "Lácteos", "unit": "kg", "price": 16.00, "qty": 0.040, "merma": 0.10},
        {"name": "Salmón ahumado", "cat": "Pescado", "unit": "kg", "price": 35.00, "qty": 0.030, "merma": 0.10},
        {"name": "Langostino cocido", "cat": "Marisco", "unit": "kg", "price": 18.00, "qty": 0.050, "merma": 0.45},
        {"name": "Croquetas (ud)", "cat": "Congelados", "unit": "ud", "price": 0.35, "qty": 3.000, "use_unit": "ud", "merma": 0.05},
        {"name": "Mini quiche (ud)", "cat": "Congelados", "unit": "ud", "price": 0.45, "qty": 2.000, "use_unit": "ud", "merma": 0.05},
        {"name": "Pan mini (ud)", "cat": "Pan/bollería", "unit": "ud", "price": 0.15, "qty": 3.000, "use_unit": "ud", "merma": 0.10},
        {"name": "Fruta variada", "cat": "Fruta", "unit": "kg", "price": 3.50, "qty": 0.060, "merma": 0.15},
        {"name": "Bebida (vino + agua)", "cat": "Bebidas/licores", "unit": "L", "price": 3.00, "qty": 0.300, "merma": 0.05},
        {"name": "Café", "cat": "Bebidas/licores", "unit": "ud", "price": 0.20, "qty": 1.000, "use_unit": "ud", "merma": 0.00},
    ]

    setup_escandallo_sheet(wb, "Cocktail 50 pax", "Catering Cocktail — Coste por Persona", cocktail_items)
    # Override food cost for catering
    ws = wb["Cocktail 50 pax"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 0.30 and isinstance(cell.value, float):
                cell.value = 0.35
                break

    # Budget sheet
    ws_budget = wb.create_sheet(title="Presupuesto")
    ws_budget.sheet_properties.tabColor = "FF4444"
    ws_budget.column_dimensions["B"].width = 40
    ws_budget.column_dimensions["C"].width = 18

    ws_budget["B2"].value = "Presupuesto de Catering"
    ws_budget["B2"].font = title_font

    rows_data = [
        (4, "Concepto", "Importe (€)", True),
        (5, "Número de comensales", 50, False),
        (6, "Coste alimentación por persona", "='Cocktail 50 pax'!I17", False),
        (7, "Coste total alimentación", "=C5*C6", False),
        (8, "Personal de servicio (camareros)", "=C5*3.50", False),
        (9, "Transporte y logística", 150.00, False),
        (10, "Material desechable / menaje", "=C5*1.20", False),
        (11, "Montaje y desmontaje", 200.00, False),
        (13, "COSTE TOTAL EVENTO", "=SUM(C7:C11)", True),
        (14, "Coste por persona", "=C13/C5", False),
        (16, "Margen objetivo (%)", 0.35, False),
        (17, "PVP POR PERSONA (sin IVA)", "=C14/(1-C16)", True),
        (18, "PVP POR PERSONA (con IVA 10%)", "=C17*1.10", True),
        (20, "PRESUPUESTO TOTAL (con IVA)", "=C18*C5", True),
    ]

    for row_num, label, value, is_total in rows_data:
        ws_budget.cell(row=row_num, column=2, value=label).border = thin_border
        cell_c = ws_budget.cell(row=row_num, column=3, value=value)
        cell_c.border = thin_border

        if row_num == 4:
            ws_budget.cell(row=row_num, column=2).font = header_font
            ws_budget.cell(row=row_num, column=2).fill = header_fill
            cell_c.font = header_font
            cell_c.fill = header_fill
        elif is_total:
            ws_budget.cell(row=row_num, column=2).font = total_font
            ws_budget.cell(row=row_num, column=2).fill = gold_fill
            cell_c.font = Font(name="Calibri", size=13, bold=True)
            cell_c.fill = gold_fill
            cell_c.number_format = '#,##0.00 €'
        elif row_num == 5:
            cell_c.fill = input_fill
            cell_c.number_format = '0'
        elif row_num == 16:
            cell_c.fill = input_fill
            cell_c.number_format = '0%'
        else:
            cell_c.number_format = '#,##0.00 €'
            if isinstance(value, (int, float)):
                cell_c.fill = input_fill

    path = os.path.join(OUTPUT_DIR, "06-catering.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 7: Cafetería/Brunch
# ══════════════════════════════════════════════════════════════
def create_cafeteria():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Cafetería y Brunch", [
        "Cómo usar esta plantilla",
        "",
        "▸ 3 recetas típicas de brunch/cafetería",
        "▸ Food cost objetivo en cafetería: 25-30%",
        "▸ Ideal para calcular precios de carta de brunch",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    recetas = [
        ("Tostada Aguacate", "Tostada de Aguacate con Huevo Poché", [
            {"name": "Pan de masa madre", "cat": "Pan/bollería", "unit": "ud", "price": 0.60, "qty": 2.000, "use_unit": "ud", "merma": 0.00},
            {"name": "Aguacate Hass", "cat": "Fruta", "unit": "ud", "price": 1.50, "qty": 1.000, "use_unit": "ud", "merma": 0.30},
            {"name": "Huevo campero", "cat": "Huevos", "unit": "docena", "price": 4.20, "qty": 0.167, "use_unit": "ud", "merma": 0.11},
            {"name": "Tomate cherry", "cat": "Verdura hoja", "unit": "kg", "price": 4.00, "qty": 0.040, "merma": 0.05},
            {"name": "Semillas mix", "cat": "Secos/granos", "unit": "kg", "price": 10.00, "qty": 0.010, "merma": 0.02},
            {"name": "AOVE", "cat": "Aceites/grasas", "unit": "L", "price": 7.50, "qty": 0.010, "merma": 0.03},
            {"name": "Sal en escamas", "cat": "Secos/granos", "unit": "kg", "price": 15.00, "qty": 0.002, "merma": 0.02},
        ]),
        ("Açaí Bowl", "Açaí Bowl con Granola y Fruta", [
            {"name": "Açaí congelado", "cat": "Congelados", "unit": "kg", "price": 22.00, "qty": 0.100, "merma": 0.07},
            {"name": "Plátano", "cat": "Fruta", "unit": "kg", "price": 1.50, "qty": 0.120, "merma": 0.25},
            {"name": "Leche de avena", "cat": "Bebidas/licores", "unit": "L", "price": 2.50, "qty": 0.080, "merma": 0.03},
            {"name": "Granola artesanal", "cat": "Secos/granos", "unit": "kg", "price": 8.00, "qty": 0.040, "merma": 0.02},
            {"name": "Arándanos", "cat": "Fruta", "unit": "kg", "price": 12.00, "qty": 0.030, "merma": 0.05},
            {"name": "Coco rallado", "cat": "Secos/granos", "unit": "kg", "price": 7.00, "qty": 0.010, "merma": 0.02},
            {"name": "Miel", "cat": "Secos/granos", "unit": "kg", "price": 10.00, "qty": 0.015, "merma": 0.03},
        ]),
        ("Eggs Benedict", "Eggs Benedict con Salmón Ahumado", [
            {"name": "English muffin", "cat": "Pan/bollería", "unit": "ud", "price": 0.40, "qty": 2.000, "use_unit": "ud", "merma": 0.00},
            {"name": "Huevo campero", "cat": "Huevos", "unit": "docena", "price": 4.20, "qty": 0.167, "use_unit": "ud", "merma": 0.11},
            {"name": "Salmón ahumado", "cat": "Pescado", "unit": "kg", "price": 35.00, "qty": 0.060, "merma": 0.05},
            {"name": "Mantequilla (hollandaise)", "cat": "Lácteos", "unit": "kg", "price": 9.00, "qty": 0.050, "merma": 0.03},
            {"name": "Limón", "cat": "Fruta", "unit": "kg", "price": 2.50, "qty": 0.010, "merma": 0.30},
            {"name": "Cebollino", "cat": "Especias/hierbas", "unit": "manojo", "price": 1.50, "qty": 0.250, "use_unit": "ud", "merma": 0.20},
        ]),
    ]

    for sheet_name, title, ingredients in recetas:
        ws, _ = setup_escandallo_sheet(wb, sheet_name, title, ingredients)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 0.30 and isinstance(cell.value, float):
                    cell.value = 0.28
                    break

    path = os.path.join(OUTPUT_DIR, "07-cafeteria-brunch.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 8: Food Truck
# ══════════════════════════════════════════════════════════════
def create_food_truck():
    wb = Workbook()
    add_instructions_sheet(wb, "Escandallo — Food Truck", [
        "Cómo usar esta plantilla",
        "",
        "▸ Recetas simplificadas con foco en velocidad y margen",
        "▸ Food cost objetivo en food truck: 28-35%",
        "▸ Los platos de food truck deben tener márgenes altos",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    recetas = [
        ("Smash Burger", "Smash Burger Doble con Queso", [
            {"name": "Carne picada (blend 80/20)", "cat": "Carne roja", "unit": "kg", "price": 9.50, "qty": 0.180, "merma": 0.15},
            {"name": "Pan brioche", "cat": "Pan/bollería", "unit": "ud", "price": 0.40, "qty": 1.000, "use_unit": "ud", "merma": 0.00},
            {"name": "Queso americano", "cat": "Lácteos", "unit": "ud", "price": 0.15, "qty": 2.000, "use_unit": "ud", "merma": 0.00},
            {"name": "Cebolla", "cat": "Verdura raíz", "unit": "kg", "price": 1.20, "qty": 0.040, "merma": 0.12},
            {"name": "Lechuga", "cat": "Verdura hoja", "unit": "kg", "price": 2.50, "qty": 0.020, "merma": 0.25},
            {"name": "Tomate", "cat": "Verdura hoja", "unit": "kg", "price": 2.50, "qty": 0.030, "merma": 0.10},
            {"name": "Salsa especial", "cat": "Secos/granos", "unit": "L", "price": 5.00, "qty": 0.025, "merma": 0.05},
            {"name": "Pepinillo", "cat": "Congelados", "unit": "kg", "price": 4.00, "qty": 0.020, "merma": 0.02},
        ]),
        ("Loaded Fries", "Loaded Fries con Pulled Pork", [
            {"name": "Patata para freír", "cat": "Verdura raíz", "unit": "kg", "price": 1.50, "qty": 0.250, "merma": 0.12},
            {"name": "Pulled pork", "cat": "Carne roja", "unit": "kg", "price": 14.00, "qty": 0.100, "merma": 0.10},
            {"name": "Queso cheddar rallado", "cat": "Lácteos", "unit": "kg", "price": 10.00, "qty": 0.040, "merma": 0.03},
            {"name": "Salsa BBQ", "cat": "Secos/granos", "unit": "L", "price": 4.00, "qty": 0.030, "merma": 0.05},
            {"name": "Cebollino", "cat": "Especias/hierbas", "unit": "manojo", "price": 1.50, "qty": 0.125, "use_unit": "ud", "merma": 0.20},
            {"name": "Aceite de girasol (fritura)", "cat": "Aceites/grasas", "unit": "L", "price": 2.00, "qty": 0.100, "merma": 0.00},
        ]),
        ("Pulled Pork Sándwich", "Pulled Pork Sándwich", [
            {"name": "Pulled pork", "cat": "Carne roja", "unit": "kg", "price": 14.00, "qty": 0.150, "merma": 0.10},
            {"name": "Pan brioche grande", "cat": "Pan/bollería", "unit": "ud", "price": 0.50, "qty": 1.000, "use_unit": "ud", "merma": 0.00},
            {"name": "Coleslaw (mix)", "cat": "Verdura hoja", "unit": "kg", "price": 3.00, "qty": 0.060, "merma": 0.10},
            {"name": "Salsa BBQ", "cat": "Secos/granos", "unit": "L", "price": 4.00, "qty": 0.030, "merma": 0.05},
            {"name": "Pepinillos", "cat": "Congelados", "unit": "kg", "price": 4.00, "qty": 0.020, "merma": 0.02},
        ]),
    ]

    for sheet_name, title, ingredients in recetas:
        ws, _ = setup_escandallo_sheet(wb, sheet_name, title, ingredients)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 0.30 and isinstance(cell.value, float):
                    cell.value = 0.30
                    break

    path = os.path.join(OUTPUT_DIR, "08-food-truck.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 9: Control de Mermas
# ══════════════════════════════════════════════════════════════
def create_control_mermas():
    wb = Workbook()
    add_instructions_sheet(wb, "Control de Mermas por Categoría", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra las mermas reales de tu establecimiento semanalmente",
        "▸ Compara con la merma estándar de la industria",
        "▸ Identifica categorías donde pierdes más de lo normal",
        "▸ La columna 'Diferencia' te muestra si estás por encima o por debajo",
        "▸ Rojo = merma real superior al estándar (alerta)",
        "▸ Verde = merma real igual o inferior al estándar (bien)",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Mermas Semanal")
    ws.sheet_properties.tabColor = GOLD

    widths = {"A": 22, "B": 14, "C": 16, "D": 16, "E": 14, "F": 14, "G": 14, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Control de Mermas — Registro Semanal"
    ws["A1"].font = title_font

    ws["A2"].value = "Semana del:"
    ws["A2"].font = subtitle_font
    ws["B2"].fill = input_fill
    ws["B2"].border = thin_border

    headers = [
        "Categoría", "Merma Estándar", "Compra Total (€)",
        "Merma Real (€)", "Merma Real (%)", "Diferencia (%)", "Estado", "Notas"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, (cat, merma) in enumerate(MERMAS.items()):
        row = 5 + i
        fill = light_row if i % 2 == 0 else alt_row

        ws.cell(row=row, column=1, value=cat).font = data_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=merma).font = data_font
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).number_format = '0%'
        ws.cell(row=row, column=2).alignment = center_align

        ws.cell(row=row, column=3).fill = input_fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).number_format = '#,##0.00 €'

        ws.cell(row=row, column=4).fill = input_fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0.00 €'

        # Merma real %
        ws.cell(row=row, column=5, value=f'=IF(C{row}=0,"",D{row}/C{row})').font = formula_font
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '0.0%'
        ws.cell(row=row, column=5).alignment = center_align

        # Diferencia
        ws.cell(row=row, column=6, value=f'=IF(E{row}="","",E{row}-B{row})').font = formula_font
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).number_format = '+0.0%;-0.0%'
        ws.cell(row=row, column=6).alignment = center_align

        # Estado
        ws.cell(row=row, column=7, value=f'=IF(E{row}="","",IF(E{row}<=B{row},"OK","ALERTA"))').font = formula_font
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).alignment = center_align

        # Notas
        ws.cell(row=row, column=8).fill = input_fill
        ws.cell(row=row, column=8).border = thin_border

    # Totals
    last_row = 5 + len(MERMAS)
    ws.cell(row=last_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=last_row, column=1).fill = gold_fill
    ws.cell(row=last_row, column=1).border = thin_border
    ws.cell(row=last_row, column=3, value=f"=SUM(C5:C{last_row-1})").font = total_font
    ws.cell(row=last_row, column=3).fill = gold_fill
    ws.cell(row=last_row, column=3).border = thin_border
    ws.cell(row=last_row, column=3).number_format = '#,##0.00 €'
    ws.cell(row=last_row, column=4, value=f"=SUM(D5:D{last_row-1})").font = total_font
    ws.cell(row=last_row, column=4).fill = gold_fill
    ws.cell(row=last_row, column=4).border = thin_border
    ws.cell(row=last_row, column=4).number_format = '#,##0.00 €'
    ws.cell(row=last_row, column=5, value=f"=IF(C{last_row}=0,\"\",D{last_row}/C{last_row})").font = total_font
    ws.cell(row=last_row, column=5).fill = gold_fill
    ws.cell(row=last_row, column=5).border = thin_border
    ws.cell(row=last_row, column=5).number_format = '0.0%'

    ws.freeze_panes = "A5"

    path = os.path.join(OUTPUT_DIR, "09-control-mermas.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 10: Calculadora de PVP
# ══════════════════════════════════════════════════════════════
def create_calculadora_pvp():
    wb = Workbook()
    add_instructions_sheet(wb, "Calculadora de PVP por Tipo de Establecimiento", [
        "Cómo usar esta plantilla",
        "",
        "▸ Introduce el coste total de tu plato en la celda verde",
        "▸ La tabla calcula automáticamente el PVP para cada tipo de establecimiento",
        "▸ Cada tipo tiene su food cost objetivo y multiplicador",
        "▸ Los PVP incluyen versión sin IVA y con IVA (10%)",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Calculadora PVP")
    ws.sheet_properties.tabColor = GOLD

    widths = {"A": 5, "B": 25, "C": 16, "D": 14, "E": 16, "F": 16, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws["B2"].value = "Calculadora de PVP Sugerido"
    ws["B2"].font = title_font

    ws["B4"].value = "COSTE TOTAL DEL PLATO (€):"
    ws["B4"].font = Font(name="Calibri", size=13, bold=True, color=DARK_BG)
    ws["C4"].value = 5.50
    ws["C4"].fill = input_fill
    ws["C4"].border = thin_border
    ws["C4"].font = Font(name="Calibri", size=16, bold=True)
    ws["C4"].number_format = '#,##0.00 €'

    headers = ["Tipo Establecimiento", "Food Cost Min", "Food Cost Max", "PVP sin IVA (€)", "PVP con IVA (€)", "Margen (€)"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, (tipo, data) in enumerate(FOOD_COST_TARGETS.items()):
        row = 7 + i
        fill = light_row if i % 2 == 0 else alt_row

        ws.cell(row=row, column=2, value=tipo).font = data_font
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=data["fc_min"]).font = data_font
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).number_format = '0%'
        ws.cell(row=row, column=3).alignment = center_align

        ws.cell(row=row, column=4, value=data["fc_max"]).font = data_font
        ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '0%'
        ws.cell(row=row, column=4).alignment = center_align

        # PVP sin IVA (using average of min/max)
        avg_fc = f"=AVERAGE(C{row},D{row})"
        ws.cell(row=row, column=5, value=f"=$C$4/{avg_fc}").font = formula_font
        ws.cell(row=row, column=5).fill = fill
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '#,##0.00 €'
        ws.cell(row=row, column=5).alignment = right_align

        # PVP con IVA
        ws.cell(row=row, column=6, value=f"=E{row}*1.10").font = formula_font
        ws.cell(row=row, column=6).fill = total_fill if i == 0 else fill
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).number_format = '#,##0.00 €'
        ws.cell(row=row, column=6).alignment = right_align

        # Margen
        ws.cell(row=row, column=7, value=f"=E{row}-$C$4").font = formula_font
        ws.cell(row=row, column=7).fill = fill
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).number_format = '#,##0.00 €'
        ws.cell(row=row, column=7).alignment = right_align

    path = os.path.join(OUTPUT_DIR, "10-calculadora-pvp.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 11: Dashboard Food Cost Mensual
# ══════════════════════════════════════════════════════════════
def create_dashboard_mensual():
    wb = Workbook()
    add_instructions_sheet(wb, "Dashboard Food Cost Mensual", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra compras y ventas mensuales en la tabla",
        "▸ El food cost % se calcula automáticamente",
        "▸ El gráfico se actualiza con los datos introducidos",
        "▸ Compara tu food cost real con el objetivo mes a mes",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Dashboard")
    ws.sheet_properties.tabColor = GOLD

    widths = {"A": 5, "B": 14, "C": 18, "D": 18, "E": 14, "F": 14, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws["B2"].value = "Dashboard Food Cost Mensual"
    ws["B2"].font = title_font

    ws["B3"].value = "Año:"
    ws["B3"].font = subtitle_font
    ws["C3"].value = 2026
    ws["C3"].fill = input_fill
    ws["C3"].border = thin_border

    ws["B4"].value = "Food Cost Objetivo (%):"
    ws["B4"].font = subtitle_font
    ws["C4"].value = 0.30
    ws["C4"].fill = input_fill
    ws["C4"].border = thin_border
    ws["C4"].number_format = '0%'

    headers = ["Mes", "Compras (€)", "Ventas (€)", "Food Cost %", "Objetivo %", "Diferencia"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Example data for first 3 months
    example_data = [
        (8500, 28000), (7800, 26500), (9200, 30000),
    ]

    for i, mes in enumerate(meses):
        row = 7 + i
        fill = light_row if i % 2 == 0 else alt_row

        ws.cell(row=row, column=2, value=mes).font = data_font
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border

        # Compras
        ws.cell(row=row, column=3).fill = input_fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).number_format = '#,##0 €'
        if i < len(example_data):
            ws.cell(row=row, column=3).value = example_data[i][0]

        # Ventas
        ws.cell(row=row, column=4).fill = input_fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0 €'
        if i < len(example_data):
            ws.cell(row=row, column=4).value = example_data[i][1]

        # Food Cost %
        ws.cell(row=row, column=5, value=f'=IF(D{row}=0,"",C{row}/D{row})').font = formula_font
        ws.cell(row=row, column=5).fill = fill
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '0.0%'
        ws.cell(row=row, column=5).alignment = center_align

        # Objetivo
        ws.cell(row=row, column=6, value=f'=$C$4').font = data_font
        ws.cell(row=row, column=6).fill = fill
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).number_format = '0.0%'
        ws.cell(row=row, column=6).alignment = center_align

        # Diferencia
        ws.cell(row=row, column=7, value=f'=IF(E{row}="","",E{row}-F{row})').font = formula_font
        ws.cell(row=row, column=7).fill = fill
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).number_format = '+0.0%;-0.0%'
        ws.cell(row=row, column=7).alignment = center_align

    # Totals row
    total_row = 19
    ws.cell(row=total_row, column=2, value="TOTAL ANUAL").font = total_font
    ws.cell(row=total_row, column=2).fill = gold_fill
    ws.cell(row=total_row, column=2).border = thin_border
    ws.cell(row=total_row, column=3, value=f"=SUM(C7:C18)").font = total_font
    ws.cell(row=total_row, column=3).fill = gold_fill
    ws.cell(row=total_row, column=3).border = thin_border
    ws.cell(row=total_row, column=3).number_format = '#,##0 €'
    ws.cell(row=total_row, column=4, value=f"=SUM(D7:D18)").font = total_font
    ws.cell(row=total_row, column=4).fill = gold_fill
    ws.cell(row=total_row, column=4).border = thin_border
    ws.cell(row=total_row, column=4).number_format = '#,##0 €'
    ws.cell(row=total_row, column=5, value=f"=IF(D{total_row}=0,\"\",C{total_row}/D{total_row})").font = total_font
    ws.cell(row=total_row, column=5).fill = gold_fill
    ws.cell(row=total_row, column=5).border = thin_border
    ws.cell(row=total_row, column=5).number_format = '0.0%'

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Food Cost Mensual vs Objetivo"
    chart.y_axis.title = "Porcentaje"
    chart.x_axis.title = "Mes"
    chart.style = 10
    chart.width = 25
    chart.height = 14

    data_fc = Reference(ws, min_col=5, min_row=6, max_row=18)
    data_obj = Reference(ws, min_col=6, min_row=6, max_row=18)
    cats = Reference(ws, min_col=2, min_row=7, max_row=18)

    chart.add_data(data_fc, titles_from_data=True)
    chart.add_data(data_obj, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.solidFill = GOLD
    chart.series[1].graphicalProperties.solidFill = "FF4444"

    ws.add_chart(chart, "B21")

    path = os.path.join(OUTPUT_DIR, "11-dashboard-food-cost-mensual.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# BONUS: Checklist Mermas + Inventario
# ══════════════════════════════════════════════════════════════
def create_bonus_mermas_inventario():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS: Checklist Mermas + Inventario", [
        "Cómo usar esta plantilla",
        "",
        "▸ Pestaña 'Inventario': registra stock y calcula consumo real vs teórico",
        "▸ Pestaña 'Checklist Mermas': control diario de desperdicios",
        "▸ Usa ambas juntas para detectar pérdidas ocultas",
        "",
        "— Kit de Escandallos Pro · AI Chef Pro · aichef.pro",
    ])

    # ── Inventario sheet ──
    ws_inv = wb.create_sheet(title="Inventario")
    ws_inv.sheet_properties.tabColor = GOLD

    inv_widths = {"A": 22, "B": 12, "C": 14, "D": 14, "E": 14, "F": 16, "G": 16, "H": 14}
    for col, w in inv_widths.items():
        ws_inv.column_dimensions[col].width = w

    ws_inv["A1"].value = "Control de Inventario — Semanal/Mensual"
    ws_inv["A1"].font = title_font
    ws_inv.merge_cells("A1:H1")

    ws_inv["A2"].value = "Periodo:"
    ws_inv["A2"].font = subtitle_font
    ws_inv["B2"].fill = input_fill
    ws_inv["B2"].border = thin_border

    inv_headers = [
        "Producto", "Ud.", "Stock Inicial", "Compras",
        "Stock Final", "Consumo Real", "Consumo Teórico", "Diferencia"
    ]
    for col_idx, h in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    example_products = [
        ("Solomillo ternera", "kg"), ("Contramuslo pollo", "kg"),
        ("Lubina", "kg"), ("Langostino", "kg"),
        ("Patata", "kg"), ("Tomate", "kg"),
        ("Lechuga", "kg"), ("Mantequilla", "kg"),
        ("Aceite oliva", "L"), ("Vino blanco", "L"),
        ("Harina", "kg"), ("Azúcar", "kg"),
        ("Huevos", "docena"), ("Nata", "L"),
        ("Chocolate", "kg"),
    ]

    for i, (prod, unit) in enumerate(example_products):
        row = 5 + i
        fill = light_row if i % 2 == 0 else alt_row

        ws_inv.cell(row=row, column=1, value=prod).font = data_font
        ws_inv.cell(row=row, column=1).fill = fill
        ws_inv.cell(row=row, column=1).border = thin_border

        ws_inv.cell(row=row, column=2, value=unit).font = data_font
        ws_inv.cell(row=row, column=2).fill = fill
        ws_inv.cell(row=row, column=2).border = thin_border
        ws_inv.cell(row=row, column=2).alignment = center_align

        for col in [3, 4, 5, 7]:  # editable columns
            ws_inv.cell(row=row, column=col).fill = input_fill
            ws_inv.cell(row=row, column=col).border = thin_border
            ws_inv.cell(row=row, column=col).number_format = '0.00'

        # Consumo real = stock inicial + compras - stock final
        ws_inv.cell(row=row, column=6, value=f"=C{row}+D{row}-E{row}").font = formula_font
        ws_inv.cell(row=row, column=6).fill = fill
        ws_inv.cell(row=row, column=6).border = thin_border
        ws_inv.cell(row=row, column=6).number_format = '0.00'

        # Diferencia = consumo real - consumo teórico
        ws_inv.cell(row=row, column=8, value=f"=F{row}-G{row}").font = formula_font
        ws_inv.cell(row=row, column=8).fill = fill
        ws_inv.cell(row=row, column=8).border = thin_border
        ws_inv.cell(row=row, column=8).number_format = '+0.00;-0.00'

    ws_inv.freeze_panes = "A5"

    # ── Checklist Mermas Diario ──
    ws_check = wb.create_sheet(title="Checklist Mermas")
    ws_check.sheet_properties.tabColor = "FF4444"

    ws_check.column_dimensions["A"].width = 14
    ws_check.column_dimensions["B"].width = 22
    ws_check.column_dimensions["C"].width = 12
    ws_check.column_dimensions["D"].width = 14
    ws_check.column_dimensions["E"].width = 30
    ws_check.column_dimensions["F"].width = 16

    ws_check["A1"].value = "Checklist de Mermas — Registro Diario"
    ws_check["A1"].font = title_font
    ws_check.merge_cells("A1:F1")

    check_headers = ["Fecha", "Producto", "Cantidad", "Motivo", "Acción correctiva", "Coste estimado (€)"]
    for col_idx, h in enumerate(check_headers, 1):
        cell = ws_check.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    motivo_validation = DataValidation(
        type="list",
        formula1='"Caducidad,Mal estado,Sobreproducción,Error de preparación,Almacenaje incorrecto,Porcionado excesivo,Otro"',
        allow_blank=True
    )
    ws_check.add_data_validation(motivo_validation)

    for i in range(30):  # 30 rows for a month
        row = 4 + i
        fill = light_row if i % 2 == 0 else alt_row
        for col in range(1, 7):
            cell = ws_check.cell(row=row, column=col)
            cell.fill = input_fill
            cell.border = thin_border
            if col == 1:
                cell.number_format = 'DD/MM/YYYY'
            if col == 6:
                cell.number_format = '#,##0.00 €'
        motivo_validation.add(ws_check.cell(row=row, column=4))

    # Total
    total_row = 34
    ws_check.cell(row=total_row, column=1, value="TOTAL MERMAS MES").font = total_font
    ws_check.cell(row=total_row, column=1).fill = gold_fill
    ws_check.cell(row=total_row, column=1).border = thin_border
    ws_check.merge_cells(f"A{total_row}:E{total_row}")
    ws_check.cell(row=total_row, column=6, value=f"=SUM(F4:F33)").font = Font(name="Calibri", size=14, bold=True)
    ws_check.cell(row=total_row, column=6).fill = gold_fill
    ws_check.cell(row=total_row, column=6).border = thin_border
    ws_check.cell(row=total_row, column=6).number_format = '#,##0.00 €'

    ws_check.freeze_panes = "A4"

    path = os.path.join(OUTPUT_DIR, "BONUS-mermas-inventario.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"\n🍳 Generando Kit de Escandallos Pro...")
    print(f"   Output: {OUTPUT_DIR}\n")

    create_escandallo_estandar()    # 01
    create_menu_degustacion()       # 02
    create_menu_del_dia()           # 03
    create_cocktails()              # 04
    create_pasteleria()             # 05
    create_catering()               # 06
    create_cafeteria()              # 07
    create_food_truck()             # 08
    create_control_mermas()         # 09
    create_calculadora_pvp()        # 10
    create_dashboard_mensual()      # 11
    create_bonus_mermas_inventario()  # BONUS

    print(f"\n✅ 12 archivos Excel generados correctamente")
    print(f"   Ubicación: {OUTPUT_DIR}")

    # List files
    files = sorted(os.listdir(OUTPUT_DIR))
    total_size = 0
    for f in files:
        size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
        total_size += size
        print(f"   {f} ({size//1024} KB)")
    print(f"   Total: {total_size//1024} KB")
