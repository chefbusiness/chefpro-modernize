#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Pastelería / Obrador.
AI Chef Pro — aichef.pro

Pre-filled recurring task checklists for artisan bakeries and pastry shops.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Output ──
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-pasteleria"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Brand Colors ──
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Zone colors
OBRADOR_COLOR = "FFF3E0"        # warm orange — production
HORNO_COLOR = "FFEBEE"          # warm red — ovens
VITRINA_COLOR = "E3F2FD"        # blue — display
ALMACEN_COLOR = "E8F5E9"        # green — storage
LIMPIEZA_COLOR = "FCE4EC"       # pink — cleaning
ADMIN_COLOR = "FFF8E1"          # yellow — admin
EVENTO_COLOR = "E0F2F1"         # teal — events
DESPACHO_COLOR = "F3E5F5"       # purple — counter/sales

ZONE_COLORS = {
    "Obrador": OBRADOR_COLOR,
    "Horno": HORNO_COLOR,
    "Vitrina": VITRINA_COLOR,
    "Almacén": ALMACEN_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Evento": EVENTO_COLOR,
    "Despacho": DESPACHO_COLOR,
    "General": LIGHT_GRAY,
}

# ── Fonts ──
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

# ── Fills ──
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

# ── Borders ──
thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

# ── Alignments ──
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80

    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)

    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font

    ws.merge_cells("A2:G2")
    ws["A2"].value = "Fecha: ___/___/______    Turno: ☐ Mañana  ☐ Tarde  ☐ Noche    Responsable turno: _________________________"
    ws["A2"].font = subtitle_font

    headers = ["☐", "Tarea", "Zona", "Responsable", "Hora Límite", "Hecha", "Firma"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    done_validation = DataValidation(type="list", formula1='"✓,—"', allow_blank=True)
    ws.add_data_validation(done_validation)

    current_row = 5

    for section_title, zone, tasks in sections:
        ws.merge_cells(f"A{current_row}:G{current_row}")
        cell = ws.cell(row=current_row, column=1, value=f"  {section_title}")
        cell.font = section_font
        zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
        section_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).fill = section_fill
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1

        for task_desc, responsible, time_limit in tasks:
            ws.cell(row=current_row, column=1, value="☐").font = checkbox_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value=task_desc).font = data_font
            ws.cell(row=current_row, column=2).alignment = left_align
            ws.cell(row=current_row, column=2).border = thin_border

            zone_fill = PatternFill(start_color=ZONE_COLORS.get(zone, LIGHT_GRAY),
                                     end_color=ZONE_COLORS.get(zone, LIGHT_GRAY), fill_type="solid")
            ws.cell(row=current_row, column=3, value=zone).font = Font(name="Calibri", size=10, color="666666")
            ws.cell(row=current_row, column=3).fill = zone_fill
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=3).border = thin_border

            ws.cell(row=current_row, column=4, value=responsible).font = data_font
            ws.cell(row=current_row, column=4).fill = input_fill
            ws.cell(row=current_row, column=4).alignment = center_align
            ws.cell(row=current_row, column=4).border = thin_border

            ws.cell(row=current_row, column=5, value=time_limit).font = data_font
            ws.cell(row=current_row, column=5).alignment = center_align
            ws.cell(row=current_row, column=5).border = thin_border

            cell_done = ws.cell(row=current_row, column=6)
            cell_done.fill = input_fill
            cell_done.alignment = center_align
            cell_done.border = thin_border
            done_validation.add(cell_done)

            ws.cell(row=current_row, column=7).fill = input_fill
            ws.cell(row=current_row, column=7).border = thin_border

            current_row += 1

        current_row += 1

    current_row += 1
    ws.merge_cells(f"A{current_row}:C{current_row}")
    ws.cell(row=current_row, column=1, value="Tareas completadas:").font = bold_font
    ws.cell(row=current_row, column=4, value=f'=COUNTIF(F5:F{current_row-2},"✓")').font = bold_font
    ws.cell(row=current_row, column=5, value="de").font = data_font
    total_tasks = sum(len(tasks) for _, _, tasks in sections)
    ws.cell(row=current_row, column=6, value=total_tasks).font = bold_font

    current_row += 2
    ws.cell(row=current_row, column=1, value="Verificado por:").font = bold_font
    ws.merge_cells(f"B{current_row}:D{current_row}")
    ws.cell(row=current_row, column=2).fill = input_fill
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=5, value="Firma:").font = bold_font
    ws.merge_cells(f"F{current_row}:G{current_row}")
    ws.cell(row=current_row, column=6).fill = input_fill
    ws.cell(row=current_row, column=6).border = thin_border

    current_row += 2
    ws.cell(row=current_row, column=1,
            value="— Kit de Tareas Recurrentes · Pastelería / Obrador · AI Chef Pro · aichef.pro").font = small_font


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 1: Apertura y Cierre
# ═══════════════════════════════════════════════════════════════════
def generate_01_apertura_cierre():
    wb = Workbook()
    add_instructions_sheet(wb, "01 — Apertura y Cierre · Pastelería / Obrador", [
        "Cómo usar esta plantilla:",
        "▸ Imprime el checklist del turno correspondiente cada día",
        "▸ El responsable de turno marca cada tarea al completarla",
        "▸ Al finalizar el turno, firma y entrega al siguiente responsable",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables — ajusta responsables y horarios",
        "▸ Añade tareas específicas de tu obrador en las filas vacías del final",
        "▸ Borra las tareas que no apliquen a tu negocio",
    ])

    create_task_sheet(wb, "Apertura Obrador", "FF8C00", "Apertura Obrador / Zona de Producción", [
        ("Encendido de Equipos", "Obrador", [
            ("Encender hornos (precalentar a temperatura según producción del día)", "Pastelero", "05:00"),
            ("Encender amasadora / batidora planetaria", "Pastelero", "05:00"),
            ("Encender cámara de fermentación controlada (verificar Tª y humedad)", "Pastelero", "05:00"),
            ("Encender temperadora de chocolate (si aplica)", "Pastelero", "05:15"),
            ("Verificar temperatura de cámaras frigoríficas (registrar)", "Pastelero", "05:00"),
            ("Encender laminadora de masa (si aplica)", "Pastelero", "05:15"),
            ("Comprobar stock de gas / electricidad (sin cortes)", "Pastelero", "05:00"),
        ]),
        ("Preparación del Obrador", "Obrador", [
            ("Sacar masas de cámara de fermentación (revisar estado)", "Pastelero", "05:00"),
            ("Sacar mantequilla/margarina para templar (20 min antes de laminar)", "Pastelero", "05:00"),
            ("Preparar mise en place de ingredientes del día", "Ayudante", "05:30"),
            ("Pesar ingredientes para las primeras masas del día", "Pastelero", "05:30"),
            ("Revisar orden de producción del día (pedidos especiales)", "Jefe Pastelero", "05:00"),
            ("Preparar moldes, bandejas y papeles de horno", "Ayudante", "05:30"),
            ("Sanitizar superficies de trabajo y mesas de mármol", "Ayudante", "05:00"),
            ("Verificar stock de materias primas críticas (harina, azúcar, huevos, nata)", "Jefe Pastelero", "05:15"),
        ]),
        ("Apertura Vitrina / Despacho", "Vitrina", [
            ("Encender iluminación y climatización de vitrinas", "Dependiente", "07:30"),
            ("Colocar productos del día anterior (verificar caducidad)", "Dependiente", "07:30"),
            ("Preparar etiquetas de precio y alérgenos", "Dependiente", "07:45"),
            ("Montar bandeja de degustación (si aplica)", "Dependiente", "07:45"),
            ("Encender caja registradora / TPV", "Dependiente", "07:50"),
            ("Verificar stock de bolsas, cajas y packaging", "Dependiente", "07:45"),
            ("Limpiar cristales de vitrina y mostrador", "Dependiente", "07:30"),
        ]),
    ])

    create_task_sheet(wb, "Cierre Obrador", "8B4513", "Cierre Obrador / Zona de Producción", [
        ("Cierre de Producción", "Obrador", [
            ("Guardar masas en cámara de fermentación (etiquetar con fecha y hora)", "Pastelero", "18:00"),
            ("Guardar pre-elaboraciones en cámara (film, etiquetar, FIFO)", "Pastelero", "18:00"),
            ("Limpiar amasadora, batidora y utensilios", "Ayudante", "18:00"),
            ("Limpiar mesas de trabajo y superficies de mármol", "Ayudante", "18:15"),
            ("Apagar hornos (o programar para madrugada si aplica)", "Pastelero", "18:30"),
            ("Limpiar interior de hornos (retirar restos, raspar solera)", "Ayudante", "18:15"),
            ("Apagar laminadora, temperadora y equipos auxiliares", "Pastelero", "18:30"),
            ("Verificar temperaturas de cámaras (registrar en hoja de control)", "Pastelero", "18:30"),
            ("Recoger y almacenar ingredientes (cerrar sacos, tapar envases)", "Ayudante", "18:15"),
        ]),
        ("Cierre Vitrina / Despacho", "Vitrina", [
            ("Retirar productos de vitrina (valorar si se pueden vender mañana)", "Dependiente", "20:00"),
            ("Guardar productos recuperables en cámara (etiquetar)", "Dependiente", "20:00"),
            ("Registrar mermas del día (productos descartados)", "Dependiente", "20:00"),
            ("Limpiar vitrinas (interior y cristales)", "Dependiente", "20:15"),
            ("Apagar iluminación de vitrinas", "Dependiente", "20:15"),
            ("Cuadrar caja / cierre de TPV", "Dependiente", "20:15"),
            ("Barrer y fregar suelo de zona de venta", "Dependiente", "20:30"),
        ]),
        ("Cierre General", "General", [
            ("Sacar basura y residuos (separar reciclaje)", "Ayudante", "18:30"),
            ("Verificar que no queda producto fuera de cámara", "Jefe Pastelero", "18:30"),
            ("Cerrar llaves de gas (si aplica)", "Jefe Pastelero", "18:30"),
            ("Activar alarma y cerrar con llave", "Jefe Pastelero", "20:30"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx"))
    print("✅ 01-apertura-cierre.xlsx")


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 2: Tareas por Partida / Área de Producción
# ═══════════════════════════════════════════════════════════════════
def generate_02_partidas():
    wb = Workbook()
    add_instructions_sheet(wb, "02 — Partidas de Producción · Pastelería / Obrador", [
        "Tareas organizadas por área de producción.",
        "▸ Cada pestaña cubre un área específica del obrador",
        "▸ Las tareas están ordenadas cronológicamente dentro del turno",
        "▸ Personaliza añadiendo las recetas y elaboraciones de tu obrador",
    ])

    create_task_sheet(wb, "Masas y Fermentación", "FF8C00", "Partida de Masas y Fermentación", [
        ("Masas de Bollería (Croissant, Brioche, Danesa)", "Obrador", [
            ("Verificar estado de masa madre / prefermento (olor, volumen, textura)", "Pastelero", "05:00"),
            ("Refrescar masa madre si es necesario (alimentar con harina y agua)", "Pastelero", "05:00"),
            ("Pesar ingredientes para masa de croissant (harina, mantequilla, levadura)", "Pastelero", "05:15"),
            ("Amasar en planetaria (1ª velocidad 4 min, 2ª velocidad 8-10 min)", "Pastelero", "05:30"),
            ("Control de temperatura de masa al salir (24-26 °C ideal)", "Pastelero", "05:45"),
            ("Reposo en bloque en cámara (mínimo 1h a 4 °C)", "Pastelero", "05:45"),
            ("Laminar con mantequilla (3 pliegues simples o 2 dobles)", "Pastelero", "07:00"),
            ("Cortar y dar forma (croissant, pain au chocolat, trenza)", "Pastelero", "07:30"),
            ("Fermentar en cámara controlada (28 °C, 75% humedad, 1.5-2h)", "Pastelero", "08:00"),
            ("Pincelar con huevo antes de hornear", "Ayudante", "09:30"),
            ("Hornear (190-200 °C, 14-18 min según tamaño)", "Pastelero", "09:45"),
        ]),
        ("Masas de Pan Artesano", "Obrador", [
            ("Verificar fermentación nocturna de masas (volumen, alveolos)", "Pastelero", "05:00"),
            ("Dividir y pesar porciones según formato (barra, hogaza, chapata)", "Pastelero", "05:15"),
            ("Bolear y dar forma (tensión superficial correcta)", "Pastelero", "05:30"),
            ("Segunda fermentación en banastones / telas (45 min - 1h)", "Pastelero", "05:45"),
            ("Greñar / cortar con cuchilla antes de hornear", "Pastelero", "06:30"),
            ("Hornear con vapor (230-250 °C, vapor primeros 10 min)", "Pastelero", "06:45"),
            ("Enfriar en rejillas (no apilar, mínimo 30 min)", "Ayudante", "07:30"),
        ]),
        ("Masas Quebradas y Bases", "Obrador", [
            ("Preparar masa sablée (mantequilla pomada + azúcar + harina)", "Pastelero", "06:00"),
            ("Preparar masa brisée para quiches/tartas saladas", "Pastelero", "06:00"),
            ("Estirar y forrar moldes de tarta", "Ayudante", "06:30"),
            ("Hornear en blanco con pesos (180 °C, 15 min)", "Pastelero", "07:00"),
            ("Retirar pesos y dorar 5 min más", "Pastelero", "07:20"),
            ("Enfriar bases antes de rellenar", "Ayudante", "07:30"),
        ]),
    ])

    create_task_sheet(wb, "Cremas y Rellenos", "E91E63", "Partida de Cremas, Rellenos y Coberturas", [
        ("Cremas Base", "Obrador", [
            ("Preparar crema pastelera (leche, yemas, azúcar, maicena, vainilla)", "Pastelero", "06:00"),
            ("Enfriar rápidamente en baño maría inverso (abatir a <4 °C en 2h)", "Pastelero", "06:30"),
            ("Preparar crema de mantequilla (merengue suizo o italiana)", "Pastelero", "06:30"),
            ("Preparar ganache de chocolate (55% para relleno, 70% para cobertura)", "Pastelero", "07:00"),
            ("Preparar crema diplomática (pastelera + nata montada)", "Pastelero", "07:30"),
            ("Preparar lemon curd / curd de frutas de temporada", "Pastelero", "07:00"),
            ("Etiquetar todas las cremas con fecha y hora de elaboración", "Ayudante", "08:00"),
        ]),
        ("Mousses y Rellenos", "Obrador", [
            ("Preparar mousse de chocolate (templar chocolate + merengue + nata)", "Pastelero", "08:00"),
            ("Preparar mousse de frutas (puré + gelatina + merengue + nata)", "Pastelero", "08:00"),
            ("Preparar compota / confit de frutas para rellenos", "Ayudante", "07:30"),
            ("Preparar pralinés y praliné de frutos secos", "Pastelero", "08:30"),
            ("Montar entremet en aros (capas de mousse + inserto + bizcocho)", "Pastelero", "09:00"),
            ("Congelar entremets montados (mínimo 4h antes de glasear)", "Pastelero", "09:30"),
        ]),
        ("Coberturas y Acabados", "Obrador", [
            ("Preparar glasé espejo (gelatina + glucosa + chocolate blanco + colorante)", "Pastelero", "10:00"),
            ("Temperar chocolate de cobertura (curva: 45°→27°→31° negro / 29° leche)", "Pastelero", "10:00"),
            ("Preparar fondant para glaseado de bollería", "Ayudante", "09:00"),
            ("Preparar merengue italiano para decoración (almíbar 121 °C + claras)", "Pastelero", "10:30"),
            ("Preparar caramelo para decoraciones (160-170 °C)", "Pastelero", "11:00"),
        ]),
    ])

    create_task_sheet(wb, "Decoración y Acabado", "9C27B0", "Partida de Decoración y Acabado Final", [
        ("Decoración de Tartas y Entremets", "Obrador", [
            ("Desmoldar entremets congelados", "Pastelero", "10:00"),
            ("Glasear con glasé espejo a 35 °C (sobre rejilla)", "Pastelero", "10:00"),
            ("Decorar con frutas frescas, flores comestibles, láminas de chocolate", "Pastelero", "10:30"),
            ("Colocar etiquetas de chocolate / logotipos", "Pastelero", "10:45"),
            ("Montar tartas por encargo (según ficha de pedido)", "Pastelero", "11:00"),
            ("Fotografiar piezas especiales para RRSS (si aplica)", "Dependiente", "11:00"),
        ]),
        ("Montaje de Vitrina", "Vitrina", [
            ("Seleccionar piezas para vitrina (variedad, colores, tamaños)", "Jefe Pastelero", "08:00"),
            ("Montar vitrina de bollería (croissants, pain au chocolat, napolitanas)", "Dependiente", "08:00"),
            ("Montar vitrina de pastelería (tartas individuales, entremets, macarons)", "Dependiente", "08:30"),
            ("Montar vitrina de pan (si aplica)", "Dependiente", "08:00"),
            ("Colocar etiquetas de alérgenos e ingredientes", "Dependiente", "08:30"),
            ("Actualizar pizarra / carta del día", "Dependiente", "08:30"),
            ("Reponer vitrina cada 2 horas durante servicio", "Dependiente", "Continuo"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "02-partidas-cocina.xlsx"))
    print("✅ 02-partidas-cocina.xlsx")


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 3: Tareas del Manager
# ═══════════════════════════════════════════════════════════════════
def generate_03_manager():
    wb = Workbook()
    add_instructions_sheet(wb, "03 — Tareas del Manager · Pastelería / Obrador", [
        "Checklists para el/la responsable de la pastelería.",
        "▸ Diario: lo que hay que revisar cada día",
        "▸ Semanal: tareas organizadas por día de la semana",
        "▸ Mensual: revisiones periódicas",
        "▸ Handover: traspaso de turno (si hay turnos)",
    ])

    create_task_sheet(wb, "Diario Manager", "FFD700", "Checklist Diario del Manager / Responsable", [
        ("Primera Hora (antes de producción)", "Admin", [
            ("Revisar pedidos especiales del día (tartas encargo, catering)", "Manager", "05:00"),
            ("Verificar asistencia del equipo y cubrir bajas", "Manager", "05:00"),
            ("Revisar stock de materias primas críticas (harina, mantequilla, huevos, nata, chocolate)", "Manager", "05:15"),
            ("Confirmar entregas de proveedores del día", "Manager", "05:15"),
            ("Revisar producción planificada vs. capacidad del equipo", "Manager", "05:30"),
        ]),
        ("Durante la Jornada", "Admin", [
            ("Supervisar calidad de producción (revisar piezas acabadas)", "Manager", "10:00"),
            ("Control de temperaturas de cámaras y vitrinas (registrar)", "Manager", "10:00"),
            ("Revisar ventas vs. producción (ajustar si sobra o falta producto)", "Manager", "12:00"),
            ("Gestionar pedidos de clientes (tartas personalizadas, eventos)", "Manager", "Continuo"),
            ("Resolver incidencias (equipo averiado, falta stock, personal)", "Manager", "Continuo"),
            ("Supervisar limpieza y orden del obrador", "Manager", "14:00"),
        ]),
        ("Cierre de Día", "Admin", [
            ("Revisar mermas del día (registrar producto descartado)", "Manager", "18:00"),
            ("Planificar producción del día siguiente", "Manager", "18:00"),
            ("Hacer pedidos a proveedores para mañana", "Manager", "18:00"),
            ("Revisar pedidos especiales de los próximos 3 días", "Manager", "18:15"),
            ("Verificar cierre correcto de obrador y despacho", "Manager", "18:30"),
        ]),
    ])

    create_task_sheet(wb, "Semanal Manager", "FF8C00", "Checklist Semanal del Manager", [
        ("Lunes — Planificación", "Admin", [
            ("Planificar producción semanal (cantidades por producto por día)", "Manager", "05:00"),
            ("Revisar pedidos especiales de la semana (tartas, eventos)", "Manager", "05:30"),
            ("Verificar stock completo de materias primas", "Manager", "06:00"),
            ("Hacer pedidos semanales a proveedores principales", "Manager", "07:00"),
        ]),
        ("Martes — Equipo", "Admin", [
            ("Revisar cuadrante de turnos de la semana", "Manager", "06:00"),
            ("Reunión breve con equipo (objetivos, novedades, feedback)", "Manager", "06:30"),
            ("Revisar formación pendiente de nuevos empleados", "Manager", "07:00"),
        ]),
        ("Miércoles — Calidad", "Admin", [
            ("Auditoría de calidad: revisar 5 piezas aleatorias de vitrina", "Manager", "10:00"),
            ("Revisar registros de temperatura de cámaras (últimos 3 días)", "Manager", "10:30"),
            ("Verificar etiquetado de alérgenos actualizado", "Manager", "11:00"),
            ("Comprobar FIFO en cámaras y almacén", "Manager", "11:30"),
        ]),
        ("Jueves — Proveedores y Compras", "Admin", [
            ("Comparar precios de ingredientes clave (harina, mantequilla, chocolate)", "Manager", "07:00"),
            ("Negociar condiciones con proveedores si hay desviaciones", "Manager", "08:00"),
            ("Verificar recepción de mercancía (pesos, calidad, caducidad)", "Manager", "Recepción"),
        ]),
        ("Viernes — Preparación Fin de Semana", "Admin", [
            ("Planificar producción extra de fin de semana (30-50% más)", "Manager", "06:00"),
            ("Verificar stock para producción del fin de semana", "Manager", "06:30"),
            ("Confirmar turnos de fin de semana con el equipo", "Manager", "07:00"),
            ("Preparar pedidos especiales del sábado/domingo", "Manager", "14:00"),
        ]),
    ])

    create_task_sheet(wb, "Mensual Manager", "8B4513", "Checklist Mensual del Manager", [
        ("Revisión Financiera", "Admin", [
            ("Revisar coste de materias primas vs. ventas (food cost objetivo: 28-35%)", "Manager", "1ª semana"),
            ("Analizar mermas del mes (objetivo: <5%)", "Manager", "1ª semana"),
            ("Revisar rentabilidad por producto (margen por pieza)", "Manager", "1ª semana"),
            ("Comparar ventas mes actual vs. mes anterior y mismo mes año anterior", "Manager", "1ª semana"),
        ]),
        ("Mantenimiento", "Admin", [
            ("Mantenimiento preventivo de hornos (limpieza profunda, calibración)", "Manager", "2ª semana"),
            ("Mantenimiento de amasadora y batidora (engrasar, revisar ganchos)", "Manager", "2ª semana"),
            ("Revisar estado de laminadora", "Manager", "2ª semana"),
            ("Mantenimiento de cámaras frigoríficas (limpieza condensador)", "Manager", "2ª semana"),
            ("Calibrar termómetros y sondas", "Manager", "2ª semana"),
        ]),
        ("Evaluación y Desarrollo", "Admin", [
            ("Evaluación informal del equipo (feedback individual)", "Manager", "3ª semana"),
            ("Revisar carta / oferta: retirar productos que no venden", "Manager", "3ª semana"),
            ("Proponer nuevas recetas / productos de temporada", "Manager", "3ª semana"),
            ("Revisar competencia (visitar 2-3 pastelerías de la zona)", "Manager", "4ª semana"),
            ("Actualizar redes sociales con fotos de nuevos productos", "Manager", "4ª semana"),
        ]),
    ])

    create_task_sheet(wb, "Handover Turno", "607D8B", "Handover / Cambio de Turno", [
        ("Traspaso de Información", "General", [
            ("Estado de producción en curso (qué está en horno, qué fermenta, qué enfría)", "Saliente", "Cambio"),
            ("Pedidos especiales pendientes (tartas encargo, catering)", "Saliente", "Cambio"),
            ("Incidencias del turno (equipo averiado, falta stock, queja cliente)", "Saliente", "Cambio"),
            ("Entregas de proveedores pendientes", "Saliente", "Cambio"),
            ("Estado de vitrina (qué se ha vendido, qué falta reponer)", "Saliente", "Cambio"),
            ("Notas para producción siguiente turno", "Saliente", "Cambio"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "03-tareas-manager.xlsx"))
    print("✅ 03-tareas-manager.xlsx")


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 4: Tareas por Perfil Profesional
# ═══════════════════════════════════════════════════════════════════
def generate_04_perfiles():
    wb = Workbook()
    add_instructions_sheet(wb, "04 — Tareas por Perfil · Pastelería / Obrador", [
        "Checklists específicos para cada perfil profesional.",
        "▸ Cada pestaña = un puesto de trabajo",
        "▸ Perfecto para onboarding de personal nuevo",
        "▸ El empleado sabe exactamente qué se espera de él/ella cada día",
    ])

    create_task_sheet(wb, "Jefe Pastelero", "FFD700", "Tareas Diarias — Jefe Pastelero / Chef Pâtissier", [
        ("Inicio de Jornada", "Obrador", [
            ("Revisar planificación de producción del día", "Jefe Pastelero", "05:00"),
            ("Asignar tareas al equipo según producción", "Jefe Pastelero", "05:00"),
            ("Verificar estado de fermentaciones nocturnas", "Jefe Pastelero", "05:00"),
            ("Controlar calidad de materias primas recibidas", "Jefe Pastelero", "Recepción"),
            ("Supervisar primeras elaboraciones del día", "Jefe Pastelero", "06:00"),
        ]),
        ("Producción", "Obrador", [
            ("Elaborar piezas premium / creativas (las más complejas)", "Jefe Pastelero", "06:00"),
            ("Supervisar puntos críticos de producción (horneado, temperado, montaje)", "Jefe Pastelero", "Continuo"),
            ("Control de calidad: revisar piezas antes de vitrina", "Jefe Pastelero", "08:00"),
            ("Desarrollar y probar nuevas recetas", "Jefe Pastelero", "14:00"),
            ("Estandarizar recetas (fichas técnicas con pesos, tiempos, temperaturas)", "Jefe Pastelero", "14:00"),
        ]),
        ("Gestión", "Admin", [
            ("Calcular costes de nuevas recetas (escandallo)", "Jefe Pastelero", "14:00"),
            ("Formar al equipo en técnicas nuevas", "Jefe Pastelero", "14:30"),
            ("Planificar producción del día siguiente", "Jefe Pastelero", "17:00"),
            ("Hacer lista de pedidos a proveedores", "Jefe Pastelero", "17:30"),
        ]),
    ])

    create_task_sheet(wb, "Pastelero", "FF8C00", "Tareas Diarias — Pastelero / Oficial", [
        ("Producción de Masas", "Obrador", [
            ("Preparar masas según planificación del día", "Pastelero", "05:00"),
            ("Controlar tiempos de amasado y temperatura de masas", "Pastelero", "05:30"),
            ("Laminar masas de hojaldre / croissant", "Pastelero", "06:00"),
            ("Dar forma a bollería (croissants, pain au chocolat, etc.)", "Pastelero", "06:30"),
            ("Controlar fermentación (tiempo, temperatura, humedad)", "Pastelero", "07:00"),
            ("Hornear según parámetros de cada masa", "Pastelero", "08:00"),
        ]),
        ("Elaboraciones", "Obrador", [
            ("Preparar cremas y rellenos del día", "Pastelero", "06:00"),
            ("Montar tartas y entremets", "Pastelero", "09:00"),
            ("Decorar piezas según estándar de la casa", "Pastelero", "10:00"),
            ("Preparar pedidos especiales (tartas encargo)", "Pastelero", "11:00"),
        ]),
        ("Limpieza y Cierre", "Obrador", [
            ("Limpiar puesto de trabajo y utensilios", "Pastelero", "17:00"),
            ("Guardar productos en cámara (etiquetar)", "Pastelero", "17:30"),
            ("Dejar preparada mise en place para el día siguiente", "Pastelero", "17:30"),
        ]),
    ])

    create_task_sheet(wb, "Ayudante", "4CAF50", "Tareas Diarias — Ayudante de Pastelería", [
        ("Soporte a Producción", "Obrador", [
            ("Pesar ingredientes según receta (preparar mise en place)", "Ayudante", "05:00"),
            ("Preparar moldes (engrasar, forrar con papel de horno)", "Ayudante", "05:30"),
            ("Ayudar en formado de bollería", "Ayudante", "06:30"),
            ("Pincelar piezas con huevo antes de hornear", "Ayudante", "08:00"),
            ("Desmoldar y enfriar piezas horneadas", "Ayudante", "09:00"),
            ("Rellenar y glasear bollería bajo supervisión", "Ayudante", "09:30"),
        ]),
        ("Limpieza y Orden", "Limpieza", [
            ("Mantener puestos de trabajo limpios durante producción", "Ayudante", "Continuo"),
            ("Lavar utensilios, bandejas y moldes", "Ayudante", "Continuo"),
            ("Barrer y fregar suelos del obrador", "Ayudante", "17:00"),
            ("Organizar almacén (FIFO, etiquetado)", "Ayudante", "16:00"),
            ("Sacar basura y reciclar envases", "Ayudante", "17:30"),
        ]),
    ])

    create_task_sheet(wb, "Dependiente Vitrina", "2196F3", "Tareas Diarias — Dependiente / Vitrina", [
        ("Apertura y Venta", "Despacho", [
            ("Montar vitrina con productos del día", "Dependiente", "08:00"),
            ("Verificar etiquetas de alérgenos y precios", "Dependiente", "08:00"),
            ("Atender clientes con amabilidad y conocimiento del producto", "Dependiente", "Continuo"),
            ("Informar sobre ingredientes y alérgenos cuando pregunten", "Dependiente", "Continuo"),
            ("Gestionar pedidos de tartas personalizadas (rellenar ficha)", "Dependiente", "Continuo"),
            ("Preparar cajas y packaging para llevar", "Dependiente", "Continuo"),
        ]),
        ("Mantenimiento y Cierre", "Despacho", [
            ("Reponer vitrina cada 2 horas", "Dependiente", "Continuo"),
            ("Mantener mostrador y zona de caja limpios", "Dependiente", "Continuo"),
            ("Registrar productos que se agotan (informar a producción)", "Dependiente", "Continuo"),
            ("Cierre de caja / cuadre de TPV", "Dependiente", "20:00"),
            ("Retirar productos de vitrina al cierre", "Dependiente", "20:00"),
            ("Limpiar vitrinas y zona de venta", "Dependiente", "20:15"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "04-tareas-perfiles.xlsx"))
    print("✅ 04-tareas-perfiles.xlsx")


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 5: Tareas Semanales y Mensuales
# ═══════════════════════════════════════════════════════════════════
def generate_05_periodicas():
    wb = Workbook()
    add_instructions_sheet(wb, "05 — Tareas Semanales y Mensuales · Pastelería / Obrador", [
        "Tareas que se realizan con frecuencia semanal o mensual.",
        "▸ Limpieza profunda por zona",
        "▸ Mantenimiento de equipos",
        "▸ Revisión FIFO y caducidades",
        "▸ Inventario periódico",
    ])

    create_task_sheet(wb, "Limpieza Semanal", "E91E63", "Limpieza Profunda Semanal por Zona", [
        ("Obrador / Zona de Producción", "Limpieza", [
            ("Limpiar interior de hornos (raspar solera, vapor, desincrustante)", "Ayudante", "Lunes"),
            ("Desmontar y limpiar amasadora (gancho, bol, protecciones)", "Ayudante", "Lunes"),
            ("Desmontar y limpiar batidora planetaria", "Ayudante", "Lunes"),
            ("Limpiar laminadora (rodillos, cintas, bandejas)", "Ayudante", "Martes"),
            ("Limpiar mesas de mármol con producto específico (sin ácidos)", "Ayudante", "Martes"),
            ("Limpiar campana extractora y filtros", "Ayudante", "Miércoles"),
            ("Limpiar paredes y azulejos del obrador", "Ayudante", "Miércoles"),
            ("Fregar suelos con desengrasante industrial", "Ayudante", "Viernes"),
        ]),
        ("Cámaras y Almacén", "Almacén", [
            ("Limpieza interior de cámaras frigoríficas", "Ayudante", "Martes"),
            ("Revisión FIFO completa (mover producto antiguo adelante)", "Pastelero", "Martes"),
            ("Verificar caducidades de todos los productos almacenados", "Pastelero", "Martes"),
            ("Limpiar estanterías de almacén seco", "Ayudante", "Jueves"),
            ("Verificar stock mínimo de ingredientes secos (harina, azúcar, cacao)", "Pastelero", "Jueves"),
            ("Revisar estado de congelados (cristalización, quemaduras de frío)", "Pastelero", "Jueves"),
        ]),
        ("Vitrinas y Despacho", "Vitrina", [
            ("Limpieza profunda de vitrinas (interior, bandejas, cristales)", "Dependiente", "Lunes"),
            ("Limpiar caja registradora y báscula", "Dependiente", "Lunes"),
            ("Limpiar cristales de entrada y escaparate", "Dependiente", "Miércoles"),
            ("Organizar y limpiar zona de packaging (cajas, bolsas, cintas)", "Dependiente", "Viernes"),
        ]),
    ])

    create_task_sheet(wb, "Mantenimiento Mensual", "FF9800", "Mantenimiento Mensual de Equipos", [
        ("Hornos", "Obrador", [
            ("Limpieza profunda de horno (pirolisis o manual con productos)", "Mantenimiento", "1ª semana"),
            ("Verificar juntas de puertas de horno (sellado, desgaste)", "Mantenimiento", "1ª semana"),
            ("Calibrar termostato de hornos (comparar con termómetro de sonda)", "Mantenimiento", "1ª semana"),
            ("Revisar sistema de vapor (boquillas, electroválvula)", "Mantenimiento", "1ª semana"),
        ]),
        ("Equipos de Frío", "Obrador", [
            ("Limpiar condensadores de cámaras frigoríficas", "Mantenimiento", "2ª semana"),
            ("Verificar temperaturas reales vs. display (sonda independiente)", "Mantenimiento", "2ª semana"),
            ("Revisar juntas de puertas de cámaras", "Mantenimiento", "2ª semana"),
            ("Comprobar desagüe de cámaras", "Mantenimiento", "2ª semana"),
        ]),
        ("Otros Equipos", "Obrador", [
            ("Engrasar engranajes de amasadora y planetaria", "Mantenimiento", "3ª semana"),
            ("Revisar cintas y rodillos de laminadora", "Mantenimiento", "3ª semana"),
            ("Verificar estado de moldes y bandejas (deformaciones, antiadherente)", "Pastelero", "3ª semana"),
            ("Calibrar báscula de precisión", "Manager", "3ª semana"),
            ("Revisar estado de boquillas, mangas pasteleras y espátulas", "Pastelero", "3ª semana"),
        ]),
    ])

    create_task_sheet(wb, "Inventario", "4CAF50", "Inventario Semanal y Control de Stock", [
        ("Ingredientes Principales", "Almacén", [
            ("Contar stock de harinas (fuerza, floja, integral, centeno)", "Pastelero", "Viernes"),
            ("Contar stock de azúcares (blanco, moreno, glass, invertido)", "Pastelero", "Viernes"),
            ("Contar stock de mantequilla y margarinas", "Pastelero", "Viernes"),
            ("Contar stock de huevos (frescos y ovoproductos)", "Pastelero", "Viernes"),
            ("Contar stock de chocolate (negro, leche, blanco, cacao)", "Pastelero", "Viernes"),
            ("Contar stock de nata y lácteos", "Pastelero", "Viernes"),
            ("Contar stock de levaduras (fresca, seca, química)", "Pastelero", "Viernes"),
        ]),
        ("Ingredientes Secundarios y Decoración", "Almacén", [
            ("Contar stock de frutos secos (almendra, avellana, pistacho)", "Ayudante", "Viernes"),
            ("Contar stock de frutas (frescas, congeladas, liofilizadas)", "Ayudante", "Viernes"),
            ("Contar stock de aromas y esencias (vainilla, café, licores)", "Ayudante", "Viernes"),
            ("Contar stock de gelatinas, pectinas y espesantes", "Ayudante", "Viernes"),
            ("Contar stock de colorantes y decoraciones", "Ayudante", "Viernes"),
            ("Contar stock de packaging (cajas, bolsas, cintas, etiquetas)", "Dependiente", "Viernes"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx"))
    print("✅ 05-tareas-semanales-mensuales.xlsx")


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 6: Eventos y Festivos
# ═══════════════════════════════════════════════════════════════════
def generate_06_eventos():
    wb = Workbook()
    add_instructions_sheet(wb, "06 — Eventos y Festivos · Pastelería / Obrador", [
        "Checklists para fechas especiales y temporadas.",
        "▸ Las pastelerías facturan hasta el 40% de su año en temporadas clave",
        "▸ Navidad, Reyes, San Valentín, Día de la Madre... cada una requiere planificación",
        "▸ Personaliza con las fechas importantes de tu zona",
    ])

    create_task_sheet(wb, "Navidad", "C62828", "Temporada Navidad (Nov-Ene)", [
        ("Planificación (Noviembre)", "Admin", [
            ("Definir catálogo de Navidad (turrones, polvorones, roscón, troncos)", "Jefe Pastelero", "1 Nov"),
            ("Calcular cantidades de producción por producto", "Manager", "1 Nov"),
            ("Hacer pedidos especiales de ingredientes (mazapán, frutas confitadas, turrón)", "Manager", "5 Nov"),
            ("Contratar personal extra si es necesario", "Manager", "10 Nov"),
            ("Diseñar packaging navideño (cajas, lazos, etiquetas)", "Manager", "10 Nov"),
            ("Publicar catálogo de Navidad en RRSS y web", "Manager", "15 Nov"),
        ]),
        ("Producción Navideña (Diciembre)", "Obrador", [
            ("Producción de turrones artesanos (2-3 semanas antes)", "Pastelero", "1-15 Dic"),
            ("Producción de polvorones y mantecados", "Pastelero", "1-15 Dic"),
            ("Preparar masas de roscón de Reyes (fermentación 24-48h)", "Pastelero", "3 Ene"),
            ("Producción de troncos de Navidad (bûche de Noël)", "Pastelero", "20-24 Dic"),
            ("Preparar cestas y lotes navideños", "Ayudante", "15-24 Dic"),
            ("Control de pedidos de clientes (lista de encargos)", "Dependiente", "Continuo"),
        ]),
        ("Día de Reyes (5-6 Enero)", "Obrador", [
            ("Hornear roscones la madrugada del 5 al 6", "Pastelero", "05-06 Ene"),
            ("Rellenar con nata, trufa, crema, cabello de ángel", "Pastelero", "06 Ene"),
            ("Decorar con fruta confitada, azúcar, corona", "Pastelero", "06 Ene"),
            ("Introducir sorpresa y haba", "Ayudante", "06 Ene"),
            ("Gestionar colas y pedidos (reforzar despacho)", "Dependiente", "06 Ene"),
        ]),
    ])

    create_task_sheet(wb, "San Valentín", "E91E63", "San Valentín (14 Febrero)", [
        ("Planificación (2 semanas antes)", "Admin", [
            ("Diseñar colección San Valentín (tartas corazón, bombones, macarons)", "Jefe Pastelero", "1 Feb"),
            ("Pedir moldes de corazón y packaging especial", "Manager", "1 Feb"),
            ("Publicar en RRSS y abrir pedidos anticipados", "Manager", "5 Feb"),
            ("Calcular producción extra (50-100% más que día normal)", "Manager", "10 Feb"),
        ]),
        ("Producción (12-14 Feb)", "Obrador", [
            ("Producción de bombones y trufas artesanas", "Pastelero", "12 Feb"),
            ("Producción de tartas individuales con forma de corazón", "Pastelero", "13 Feb"),
            ("Producción de macarons con sabores especiales (rosa, frambuesa, pasión)", "Pastelero", "13 Feb"),
            ("Montar cajas regalo con selección de piezas", "Ayudante", "13 Feb"),
            ("Preparar tartas de encargo personalizadas", "Pastelero", "14 Feb"),
            ("Reforzar despacho para pico de ventas (14 Feb tarde)", "Dependiente", "14 Feb"),
        ]),
    ])

    create_task_sheet(wb, "Semana Santa", "FF6F00", "Semana Santa", [
        ("Planificación", "Admin", [
            ("Definir producción de Semana Santa (torrijas, monas, huevos de Pascua)", "Jefe Pastelero", "4 sem antes"),
            ("Pedir moldes de huevos de Pascua y colorantes", "Manager", "3 sem antes"),
            ("Producción de figuras de chocolate (conejos, huevos, campanas)", "Pastelero", "2 sem antes"),
        ]),
        ("Producción", "Obrador", [
            ("Elaborar torrijas (remojo, fritura, almíbar/canela)", "Pastelero", "Jue-Vie Santo"),
            ("Elaborar monas de Pascua (bizcocho + decoración chocolate)", "Pastelero", "Lun-Mié"),
            ("Producir huevos de chocolate rellenos", "Pastelero", "Lun-Mié"),
            ("Montar escaparate temático de Semana Santa", "Dependiente", "Dom Ramos"),
            ("Gestionar pedidos de encargo (monas, huevos personalizados)", "Dependiente", "Continuo"),
        ]),
    ])

    create_task_sheet(wb, "Día Madre-Padre", "9C27B0", "Día de la Madre / Día del Padre", [
        ("Preparación", "Admin", [
            ("Diseñar tartas especiales para la ocasión", "Jefe Pastelero", "2 sem antes"),
            ("Abrir pedidos de tartas personalizadas (mensaje, dedicatoria)", "Manager", "2 sem antes"),
            ("Preparar packaging especial", "Manager", "1 sem antes"),
        ]),
        ("Producción y Venta", "Obrador", [
            ("Producción extra de tartas (incremento 80-120%)", "Pastelero", "Día antes"),
            ("Preparar tartas con dedicatoria personalizada", "Pastelero", "Día"),
            ("Reforzar despacho (turno doble si es necesario)", "Dependiente", "Día"),
            ("Ofrecer servicio de entrega a domicilio (si aplica)", "Manager", "Día"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "06-eventos-festivos.xlsx"))
    print("✅ 06-eventos-festivos.xlsx")


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 7: Plantilla Personalizable
# ═══════════════════════════════════════════════════════════════════
def generate_07_personalizable():
    wb = Workbook()
    add_instructions_sheet(wb, "07 — Plantilla Personalizable · Pastelería / Obrador", [
        "Plantillas en blanco para crear tus propias listas de tareas.",
        "",
        "Cómo usar:",
        "▸ Copia la pestaña que más se ajuste a lo que necesitas",
        "▸ Rellena con las tareas específicas de tu pastelería",
        "▸ Usa los colores de zona para diferenciar áreas",
        "▸ Imprime en A4 y distribuye al equipo",
    ])

    # Blank template by time slot
    create_task_sheet(wb, "Por Franja Horaria", "FFD700", "Plantilla en Blanco — Por Franja Horaria", [
        ("Madrugada / Inicio Producción (04:00 - 07:00)", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Mañana / Producción Principal (07:00 - 12:00)", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Tarde / Producción + Venta (12:00 - 17:00)", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Cierre (17:00 - 20:00)", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
    ])

    # Blank template by zone
    create_task_sheet(wb, "Por Zona", "FF8C00", "Plantilla en Blanco — Por Zona de Trabajo", [
        ("Obrador", "Obrador", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Horno", "Horno", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Vitrina / Despacho", "Vitrina", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Almacén / Cámaras", "Almacén", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
    ])

    # Blank template by role
    create_task_sheet(wb, "Por Perfil", "9C27B0", "Plantilla en Blanco — Por Perfil Profesional", [
        ("Jefe Pastelero / Chef Pâtissier", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Pastelero / Oficial", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Ayudante de Pastelería", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
        ("Dependiente / Vitrina", "General", [
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
            ("", "", ""),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "07-plantilla-personalizable.xlsx"))
    print("✅ 07-plantilla-personalizable.xlsx")


# ═══════════════════════════════════════════════════════════════════
# BONUS 1: Briefing de Servicio
# ═══════════════════════════════════════════════════════════════════
def generate_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 01 — Briefing de Servicio · Pastelería", [
        "Plantilla para la reunión diaria del equipo antes de abrir.",
        "▸ 5 minutos que mejoran la comunicación del equipo",
        "▸ Repasa producción del día, encargos especiales, alérgenos",
        "▸ Imprime una copia cada día y pégala en el obrador",
    ])

    create_task_sheet(wb, "Briefing Diario", "FFD700", "Briefing Pre-Servicio — Pastelería / Obrador", [
        ("Producción del Día", "Obrador", [
            ("Productos planificados: bollería (croissants, napolitanas, ensaimadas...)", "Jefe Pastelero", "05:00"),
            ("Productos planificados: pastelería (tartas, entremets, individuales...)", "Jefe Pastelero", "05:00"),
            ("Productos planificados: pan (si aplica)", "Jefe Pastelero", "05:00"),
            ("Cantidades especiales (evento, festivo, fin de semana)", "Jefe Pastelero", "05:00"),
        ]),
        ("Pedidos y Encargos Especiales", "Admin", [
            ("Tartas de encargo: nombre cliente, hora recogida, detalles", "Manager", "05:00"),
            ("Pedidos de catering / lotes", "Manager", "05:00"),
            ("Entregas a domicilio programadas", "Manager", "05:00"),
        ]),
        ("Alertas del Día", "General", [
            ("Alérgenos: productos con alérgenos especiales hoy", "Jefe Pastelero", "05:00"),
            ("Producto nuevo en vitrina (ingredientes, precio, historia)", "Jefe Pastelero", "05:00"),
            ("Promociones activas", "Manager", "05:00"),
            ("Incidencias de ayer / notas del turno anterior", "Manager", "05:00"),
        ]),
        ("Equipo del Turno", "General", [
            ("Quién trabaja hoy y en qué puesto", "Manager", "05:00"),
            ("Personal nuevo / en formación (asignar tutor)", "Manager", "05:00"),
            ("Horarios especiales de hoy", "Manager", "05:00"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))
    print("✅ BONUS-01-briefing-servicio.xlsx")


# ═══════════════════════════════════════════════════════════════════
# BONUS 2: Calendario Anual de Tareas
# ═══════════════════════════════════════════════════════════════════
def generate_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 02 — Calendario Anual · Pastelería", [
        "Fechas clave para pastelerías con tareas asociadas.",
        "▸ Cada fecha incluye qué producir y con cuánta antelación",
        "▸ Añade las fechas locales de tu zona (fiestas patronales, ferias)",
        "▸ Usa como base para planificar compras y producción extra",
    ])

    ws = wb.create_sheet(title="Calendario Anual")
    ws.sheet_properties.tabColor = GOLD

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Calendario Anual de Fechas Clave — Pastelería / Obrador"
    ws["A1"].font = title_font

    headers = ["#", "Fecha", "Evento", "Producción Especial", "Antelación", "Notas"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    events = [
        ("1", "6 Enero", "Día de Reyes", "Roscón de Reyes (nata, trufa, crema)", "2 semanas", "Pico máximo del año para pastelerías"),
        ("2", "14 Febrero", "San Valentín", "Tartas corazón, bombones, macarons rosas", "2 semanas", "Packaging especial, mensajes personalizados"),
        ("3", "19 Marzo", "Día del Padre", "Tartas temáticas, pasteles de chocolate", "1 semana", "Segundo pico fuerte del Q1"),
        ("4", "Mar-Abr", "Semana Santa", "Torrijas, monas de Pascua, huevos chocolate", "3 semanas", "Producción intensiva 7-10 días"),
        ("5", "1er dom. Mayo", "Día de la Madre", "Tartas con dedicatoria, ramos de macarons", "2 semanas", "Pico muy fuerte — reforzar despacho"),
        ("6", "Junio", "Comuniones", "Tartas personalizadas, candy bars, galletas decoradas", "4 semanas", "Pedidos con mucha antelación"),
        ("7", "Jun-Sep", "Temporada Verano", "Helados artesanos, tartas frías, granizados", "4 semanas", "Adaptar vitrina a productos fríos"),
        ("8", "31 Octubre", "Halloween", "Galletas decoradas, cupcakes temáticos, tartas terror", "2 semanas", "Decoración de escaparate"),
        ("9", "1 Noviembre", "Todos los Santos", "Buñuelos de viento, huesos de santo, panellets", "2 semanas", "Muy fuerte en zonas tradicionales"),
        ("10", "25 Nov", "Black Friday", "Promociones en lotes, cajas regalo descuento", "1 semana", "Oportunidad de venta online"),
        ("11", "Diciembre", "Navidad Completa", "Turrones, polvorones, troncos, mazapanes", "6 semanas", "Mayor temporada — planificar personal extra"),
        ("12", "31 Diciembre", "Nochevieja", "Tartas para cena, petit fours, macarons", "1 semana", "Pedidos de restaurantes"),
        ("13", "Variable", "Fiestas Locales", "Dulces tradicionales de tu zona", "3 semanas", "AÑADE TUS FECHAS LOCALES"),
        ("14", "Variable", "Bodas (temporada)", "Tartas nupciales, candy bars, detalles invitados", "2-3 meses", "Abril-Octubre pico de bodas"),
        ("15", "Variable", "Eventos Corporativos", "Desayunos empresa, coffee break, cajas regalo", "2 semanas", "Todo el año, pico en Navidad"),
        ("16", "Variable", "Cumpleaños", "Tartas personalizadas por encargo", "3-7 días", "Ingresos constantes todo el año"),
        ("17", "Variable", "Bautizos/Comuniones/Aniversarios", "Tartas temáticas, galletas personalizadas", "3-4 semanas", "Fichas de pedido detalladas"),
    ]

    for i, (num, fecha, evento, produccion, antelacion, notas) in enumerate(events, start=4):
        ws.cell(row=i, column=1, value=num).font = data_font
        ws.cell(row=i, column=1).alignment = center_align
        ws.cell(row=i, column=1).border = thin_border

        ws.cell(row=i, column=2, value=fecha).font = bold_font
        ws.cell(row=i, column=2).alignment = center_align
        ws.cell(row=i, column=2).border = thin_border

        ws.cell(row=i, column=3, value=evento).font = bold_font
        ws.cell(row=i, column=3).border = thin_border

        ws.cell(row=i, column=4, value=produccion).font = data_font
        ws.cell(row=i, column=4).alignment = left_align
        ws.cell(row=i, column=4).border = thin_border

        ws.cell(row=i, column=5, value=antelacion).font = data_font
        ws.cell(row=i, column=5).alignment = center_align
        ws.cell(row=i, column=5).border = thin_border

        ws.cell(row=i, column=6, value=notas).font = Font(name="Calibri", size=10, color="666666")
        ws.cell(row=i, column=6).alignment = left_align
        ws.cell(row=i, column=6).border = thin_border

        if "AÑADE" in notas:
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = input_fill

    footer_row = len(events) + 5
    ws.cell(row=footer_row, column=1,
            value="— Kit de Tareas Recurrentes · Pastelería / Obrador · AI Chef Pro · aichef.pro").font = small_font

    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx"))
    print("✅ BONUS-02-calendario-anual-tareas.xlsx")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n🍰 Generando Kit de Tareas Recurrentes — Pastelería / Obrador\n")
    generate_01_apertura_cierre()
    generate_02_partidas()
    generate_03_manager()
    generate_04_perfiles()
    generate_05_periodicas()
    generate_06_eventos()
    generate_07_personalizable()
    generate_bonus_01()
    generate_bonus_02()
    print(f"\n✅ 9 archivos generados en {OUTPUT_DIR}\n")
