#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Restaurante Creativo / De Autor (11 archivos).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-restaurante-creativo"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Zone colors for Restaurante Creativo
COCINA_COLOR = "FFF3E0"      # Orange - main kitchen
ID_COLOR = "E8EAF6"          # Indigo - R&D/development
SALA_COLOR = "E3F2FD"        # Blue - front of house
BODEGA_COLOR = "F3E5F5"      # Purple - wine/cellar
PASTELERIA_COLOR = "FCE4EC"  # Pink - pastry
LIMPIEZA_COLOR = "EFEBE9"    # Brown - cleaning
ADMIN_COLOR = "FFF8E1"       # Amber - admin
EVENTOS_COLOR = "E0F7FA"     # Cyan - events
CAMARA_COLOR = "E8F5E9"      # Green - cold storage
RRSS_COLOR = "F1F8E9"        # Light green - social media

ZONE_COLORS = {
    "Cocina": COCINA_COLOR,
    "Prep": COCINA_COLOR,
    "Mise en place": COCINA_COLOR,
    "Producción": COCINA_COLOR,
    "Emplatado": COCINA_COLOR,
    "Pase": COCINA_COLOR,
    "Caliente": COCINA_COLOR,
    "Frío": COCINA_COLOR,
    "I+D": ID_COLOR,
    "Desarrollo": ID_COLOR,
    "Fichas": ID_COLOR,
    "Técnica": ID_COLOR,
    "Innovación": ID_COLOR,
    "Sala": SALA_COLOR,
    "Maître": SALA_COLOR,
    "Servicio": SALA_COLOR,
    "Storytelling": SALA_COLOR,
    "Camarero": SALA_COLOR,
    "Bodega": BODEGA_COLOR,
    "Sumiller": BODEGA_COLOR,
    "Maridaje": BODEGA_COLOR,
    "Vinos": BODEGA_COLOR,
    "Pastelería": PASTELERIA_COLOR,
    "Postres": PASTELERIA_COLOR,
    "Petit fours": PASTELERIA_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Post-servicio": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Finanzas": ADMIN_COLOR,
    "Food cost": ADMIN_COLOR,
    "RRHH": ADMIN_COLOR,
    "Eventos": EVENTOS_COLOR,
    "Chef's table": EVENTOS_COLOR,
    "Pop-up": EVENTOS_COLOR,
    "Cámara": CAMARA_COLOR,
    "Stock": CAMARA_COLOR,
    "Proveedores": CAMARA_COLOR,
    "RRSS": RRSS_COLOR,
    "Fotografía": RRSS_COLOR,
    "Prensa": RRSS_COLOR,
    "Marketing": RRSS_COLOR,
    "General": LIGHT_GRAY,
}

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

PRODUCT_SUBTITLE = "AI Chef Pro · aichef.pro — Kit de Tareas: Restaurante Creativo / De Autor"


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = PRODUCT_SUBTITLE
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for section_title, tasks in sections:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for idx, (task, zone) in enumerate(tasks, 1):
            zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")

            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=zone).font = data_font
            ws.cell(row=row, column=3).fill = zone_fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align

            check_cell = ws.cell(row=row, column=5)
            check_cell.font = checkbox_font
            check_cell.alignment = center_align
            check_cell.border = thin_border
            dv.add(check_cell)

            ws.cell(row=row, column=6).fill = input_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


def create_blank_template_sheet(wb, sheet_name, tab_color, title, col_headers):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = PRODUCT_SUBTITLE
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r in range(5, 30):
        for c in range(1, len(col_headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c == 1:
                cell.alignment = center_align
            elif c == 5:
                cell.font = checkbox_font
                cell.alignment = center_align
                dv.add(cell)
            elif c in (4, 6):
                cell.fill = input_fill
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.merge_cells(f"A31:G31")
    ws[f"A31"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A31"].font = small_font


def create_calendar_sheet(wb, sheet_name, tab_color, title, months_data):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:F2")
    ws["A2"].value = PRODUCT_SUBTITLE
    ws["A2"].font = subtitle_font

    row = 4
    cal_headers = ["Mes", "Acciones clave / Fechas señaladas", "Producto estrella", "Demanda", "✓", "Notas"]
    for col_idx, h in enumerate(cal_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    ws.add_data_validation(dv)

    for month, actions, product, demand in months_data:
        ws.cell(row=row, column=1, value=month).font = bold_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=actions).font = data_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=product).font = data_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = thin_border

        demand_color = COCINA_COLOR if demand == "Alta" else SALA_COLOR if demand == "Baja" else EVENTOS_COLOR
        ws.cell(row=row, column=4, value=demand).font = data_font
        ws.cell(row=row, column=4).fill = PatternFill(start_color=demand_color, end_color=demand_color, fill_type="solid")
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border

        check = ws.cell(row=row, column=5)
        check.font = checkbox_font
        check.alignment = center_align
        check.border = thin_border
        dv.add(check)

        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = left_align

        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


# ═══════════════════════════════════════════════════════════
# 01 — APERTURA Y CIERRE
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Apertura y Cierre — Restaurante Creativo", [
        "Cómo usar esta plantilla:",
        "▸ Imprime diariamente o usa en tablet para cada turno.",
        "▸ El responsable de turno marca cada tarea al completarla.",
        "▸ Revisa las fermentaciones activas y registra pH.",
        "▸ Verifica que todo el equipamiento de vanguardia esté operativo.",
        "",
        "Zonas de color:",
        "▸ Naranja = Cocina principal",
        "▸ Azul = Sala / Front of house",
        "▸ Ámbar = Administración",
        "▸ Marrón = Limpieza",
    ])
    create_task_sheet(wb, "Apertura AM", "FF9800", "Checklist de Apertura AM", [
        ("🌅 APERTURA — EQUIPAMIENTO VANGUARDIA", [
            ("Encender baños térmicos / Roner / equipos sous-vide", "Cocina"),
            ("Verificar temperatura del Roner (ajustar según prod. del día)", "Cocina"),
            ("Encender deshidratador y verificar bandejas en proceso", "Cocina"),
            ("Revisar fermentaciones activas: medir pH y registrar", "I+D"),
            ("Verificar cámaras de maduración (temp., humedad)", "Cámara"),
            ("Encender Pacojet / máquina de helados (si hay producción)", "Pastelería"),
            ("Encender sifones y cargar N2O según necesidad del menú", "Cocina"),
            ("Verificar estado del deshidratador: texturas, tiempos restantes", "Cocina"),
        ]),
        ("🔥 APERTURA — BASES Y MISE EN PLACE", [
            ("Verificar stock de bases: caldos, glacés, demi-glace", "Cocina"),
            ("Verificar aceites infusionados, vinagres fermentados", "I+D"),
            ("Comprobar stock de hidrocoloides y texturizantes", "I+D"),
            ("Revisar pre-elaboraciones en curso (sous-vide 24-72h)", "Cocina"),
            ("Montar estaciones de partida por cada pase del menú", "Mise en place"),
            ("Verificar plating bible actualizada y visible en cada pase", "Emplatado"),
            ("Comprobar stock de microbrotes, flores comestibles", "Cocina"),
            ("Verificar cortes y porciones del día (0,01g precisión)", "Cocina"),
        ]),
        ("🍽️ APERTURA — SALA Y BRIEFING", [
            ("Climatización sala: temperatura, aroma, humedad", "Sala"),
            ("Iluminación: verificar escenas de luz por turno", "Sala"),
            ("Música ambiente: playlist seleccionada para el servicio", "Sala"),
            ("Montar mesas: mantelería, vajilla de autor, cubertería", "Sala"),
            ("Verificar copas de maridaje preparadas por mesa", "Sumiller"),
            ("Briefing cocina: menú del día, cambios, alérgenos reservas", "Cocina"),
            ("Briefing sala: storytelling de cada pase, VIPs, alérgenos", "Maître"),
            ("Comprobar reservas del día y mapa de sala actualizado", "Maître"),
        ]),
    ])
    create_task_sheet(wb, "Cierre PM", "FF9800", "Checklist de Cierre PM", [
        ("🌙 CIERRE — REGISTROS Y TEMPERATURAS", [
            ("Registrar temperaturas de todas las cámaras frigoríficas", "Cámara"),
            ("Registrar temperaturas de cámaras de maduración", "Cámara"),
            ("Registrar pH de fermentaciones activas (nocturno)", "I+D"),
            ("Registrar estado de sous-vide largo recorrido (tiempo restante)", "Cocina"),
            ("Etiquetar TODAS las pre-elaboraciones (fecha, hora, contenido)", "Cocina"),
        ]),
        ("🧹 CIERRE — LIMPIEZA Y EQUIPAMIENTO", [
            ("Limpieza profunda de sifones, boquillas y cargadores", "Limpieza"),
            ("Limpieza de pinzas de emplatado, biberones, stencils", "Limpieza"),
            ("Apagar Roner / baños térmicos (excepto cocciones nocturnas)", "Cocina"),
            ("Apagar deshidratador (excepto procesos en curso >12h)", "Cocina"),
            ("Limpieza y secado de Pacojet / máquina de helados", "Limpieza"),
            ("Limpieza de sopletes, ahumadores, equipamiento especial", "Limpieza"),
            ("Limpiar y organizar estación de emplatado", "Limpieza"),
            ("Verificar que fermentaciones nocturnas están selladas", "I+D"),
        ]),
        ("📋 CIERRE — ADMINISTRACIÓN Y FEEDBACK", [
            ("Cierre de caja y registro de ventas del día", "Admin"),
            ("Registrar covers servidos (almuerzos + cenas)", "Admin"),
            ("Anotar incidencias de servicio", "Admin"),
            ("Registrar feedback del chef sobre platos del día", "I+D"),
            ("Anotar platos con mejor/peor aceptación de sala", "Sala"),
            ("Registrar mermas significativas del servicio", "Admin"),
            ("Confirmar reservas y alérgenos para mañana", "Maître"),
            ("Activar alarma y cerrar instalaciones", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx"))
    print("✅ 01-apertura-cierre.xlsx")


# ═══════════════════════════════════════════════════════════
# 02 — MISE EN PLACE MENÚ DEGUSTACIÓN
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Mise en Place — Menú Degustación", [
        "Cómo usar esta plantilla:",
        "▸ Organiza la mise en place por pases del menú degustación.",
        "▸ Distingue pre-elaboraciones por tiempo de ejecución.",
        "▸ Controla cada componente del emplatado con la plating bible.",
        "▸ Adapta la hoja según el número de pases de tu menú (7-15).",
        "",
        "Tiempos de pre-elaboración:",
        "▸ Largo recorrido (24-72h): caldos, fermentaciones, marinados",
        "▸ Medio (6-12h): sous-vide, infusiones, gelificaciones",
        "▸ Del día (2-4h): espumas, salsas calientes, cortes frescos",
        "▸ Minuto cero: microbrotes, flores, nieves, crujientes",
    ])
    create_task_sheet(wb, "Pre-elab Largo", "3F51B5", "Pre-elaboraciones Largo Recorrido (24-72h)", [
        ("🕐 CALDOS, FONDOS Y GLACÉS", [
            ("Fondo oscuro de ternera (48-72h, reducción lenta)", "Cocina"),
            ("Caldo dashi de autor (kombu + bonito + setas deshidratadas)", "Cocina"),
            ("Glacé de ave (24h reducción)", "Cocina"),
            ("Consomé clarificado (24h frío o 12h caliente)", "Cocina"),
            ("Fumet concentrado de marisco (24h)", "Cocina"),
        ]),
        ("🧪 FERMENTACIONES Y MARINADOS", [
            ("Fermentación láctica de vegetales (medir pH diario)", "I+D"),
            ("Garum / salsa fermentada en proceso (registrar día)", "I+D"),
            ("Koji: verificar estado del cultivo, temp. y humedad", "I+D"),
            ("Marinados largos (carnes/pescados en aceite, sal, especias)", "Cocina"),
            ("Encurtidos de temporada (verificar acidez)", "I+D"),
            ("Vinagres de fruta en proceso (madre activa)", "I+D"),
        ]),
        ("♨️ SOUS-VIDE Y MADURACIONES", [
            ("Sous-vide largo: piezas de carne (48-72h a baja temp.)", "Cocina"),
            ("Sous-vide medio: pescados y mariscos (6-12h)", "Cocina"),
            ("Confitados lentos (pato, cerdo, pulpo)", "Cocina"),
            ("Maduraciones controladas en cámara (registrar día)", "Cámara"),
            ("Infusiones en frío de aceites y mantequillas (24-48h)", "I+D"),
        ]),
    ])
    create_task_sheet(wb, "Pre-elab Medio", "3F51B5", "Pre-elaboraciones Medio Recorrido (6-12h)", [
        ("⏳ SOUS-VIDE E INFUSIONES (6-12h)", [
            ("Huevo sous-vide (63°C / 75 min o según receta)", "Cocina"),
            ("Vegetales sous-vide para texturas controladas", "Cocina"),
            ("Infusiones calientes de hierbas para salsas", "Cocina"),
            ("Gelificaciones en frío (agar, gellan, kappa)", "I+D"),
            ("Esfericaciones básicas e inversas (preparar baños)", "I+D"),
            ("Cremas y velouté con infusión lenta", "Cocina"),
        ]),
        ("❄️ HELADOS Y SORBETES", [
            ("Base de helado de autor (madurar 12h en frío)", "Pastelería"),
            ("Sorbetes de fruta de temporada (madurar base)", "Pastelería"),
            ("Granizados / nieves para emplatado (preparar base)", "Pastelería"),
            ("Helado salado (queso, aceite de oliva, foie)", "Pastelería"),
        ]),
    ])
    create_task_sheet(wb, "Pre-elab Día", "3F51B5", "Pre-elaboraciones del Día (2-4h)", [
        ("🔥 SALSAS Y ESPUMAS", [
            ("Espumas calientes en sifón (verificar carga N2O)", "Cocina"),
            ("Espumas frías (montar y cargar, refrigerar)", "Cocina"),
            ("Salsas calientes del menú (reducir al momento)", "Cocina"),
            ("Aires / bubble (lecitina de soja, xantana)", "I+D"),
            ("Vinagretas de autor (emulsiones estables)", "Cocina"),
        ]),
        ("🔪 CORTES Y PORCIONES", [
            ("Cortes de proteína principal (porcionar al gramo)", "Cocina"),
            ("Cortes de pescado/marisco (técnica, precisión)", "Cocina"),
            ("Torneado y cortes de vegetales de autor", "Cocina"),
            ("Láminas y carpaccios (grosor uniforme)", "Frío"),
            ("Tartares y ceviches (cortar en el momento, no antes de 2h)", "Frío"),
        ]),
        ("🌿 EMPLATADO MINUTO CERO", [
            ("Verificar stock microbrotes frescos (perejil, shiso, rábano)", "Cocina"),
            ("Verificar flores comestibles (frescura, color)", "Cocina"),
            ("Preparar nieves / granizados al momento", "Pastelería"),
            ("Preparar crujientes de última hora (tempura, chip)", "Cocina"),
            ("Montar biberones con salsas para emplatado", "Emplatado"),
            ("Verificar stencils, pinzas, cucharas de autor limpias", "Emplatado"),
        ]),
    ])
    create_task_sheet(wb, "Plating Bible", "3F51B5", "Control Plating Bible por Pase", [
        ("📖 CHECKLIST POR PASE (adaptar a tu menú 7-15 pases)", [
            ("Pase 1 — Snack: foto referencia visible, peso verificado", "Emplatado"),
            ("Pase 1 — Snack: componentes en posición correcta", "Emplatado"),
            ("Pase 2 — Aperitivo: foto referencia visible, peso verificado", "Emplatado"),
            ("Pase 2 — Aperitivo: componentes en posición correcta", "Emplatado"),
            ("Pase 3 — Entrante frío: foto referencia, peso, posición", "Emplatado"),
            ("Pase 4 — Entrante caliente: foto referencia, peso, posición", "Emplatado"),
            ("Pase 5 — Pescado: foto referencia, peso, posición", "Emplatado"),
            ("Pase 6 — Carne: foto referencia, peso, posición", "Emplatado"),
            ("Pase 7 — Pre-postre: foto referencia, peso, posición", "Emplatado"),
            ("Pase 8 — Postre: foto referencia, peso, posición", "Emplatado"),
            ("Pase 9 — Petit fours: foto referencia, peso, posición", "Emplatado"),
            ("Verificar vajilla de autor seleccionada por pase", "Emplatado"),
            ("Verificar cubertería especial por pase (si aplica)", "Sala"),
        ]),
        ("🛠️ HERRAMIENTAS DE EMPLATADO POR PASE", [
            ("Pinzas de precisión limpias y en estación", "Emplatado"),
            ("Biberones cargados con salsas correspondientes", "Emplatado"),
            ("Stencils / plantillas de decoración preparados", "Emplatado"),
            ("Sopletes cargados (para caramelizar, dorar)", "Emplatado"),
            ("Cucharas de autor (quenelle, rayas) limpias", "Emplatado"),
            ("Ahumador / campana de humo preparado (si aplica)", "Emplatado"),
            ("Ralladores (trufa, cítricos, especias) en estación", "Emplatado"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "02-mise-en-place-degustacion.xlsx"))
    print("✅ 02-mise-en-place-degustacion.xlsx")


# ═══════════════════════════════════════════════════════════
# 03 — I+D Y DESARROLLO DE MENÚ
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · I+D y Desarrollo de Menú", [
        "Cómo usar esta plantilla:",
        "▸ Gestiona el ciclo completo de innovación culinaria.",
        "▸ Documenta cada prueba con ficha completa (3 iteraciones mín.).",
        "▸ Evalúa viabilidad operativa antes de incluir en carta.",
        "▸ Mantén un calendario estacional con 4 cambios de carta/año.",
        "",
        "Protocolo I+D:",
        "▸ Cada plato nuevo: mínimo 3 iteraciones con foto y feedback.",
        "▸ Hidrocoloides: documentar al 0,01g para replicabilidad.",
        "▸ Registrar TODAS las técnicas probadas, incluidas las descartadas.",
    ])
    create_task_sheet(wb, "Sesiones I+D", "3F51B5", "Calendario y Gestión de Sesiones I+D", [
        ("🧪 CALENDARIO SESIONES I+D", [
            ("Definir día/horario fijo semanal para sesiones I+D", "I+D"),
            ("Reservar ingredientes específicos para I+D (no tocar servicio)", "I+D"),
            ("Asignar equipo para sesión I+D (chef + sous-chef mínimo)", "I+D"),
            ("Definir objetivo de la sesión: concepto, técnica o ingrediente", "I+D"),
            ("Preparar ficha de investigación previa (referencias, inspiración)", "I+D"),
            ("Reservar equipamiento especial para la sesión", "I+D"),
        ]),
        ("📝 FICHA NUEVO PLATO", [
            ("Nombre provisional / concepto del plato", "I+D"),
            ("Inspiración y referencia (origen, temporada, emoción)", "I+D"),
            ("Lista de ingredientes principales y secundarios", "I+D"),
            ("Técnica/s principal/es a emplear", "Técnica"),
            ("Coste estimado por ración (ingredientes)", "Food cost"),
            ("Tiempo de ejecución estimado en servicio", "I+D"),
            ("Dificultad de emplatado (1-5 escala)", "Emplatado"),
            ("Compatibilidad con alérgenos comunes (14 UE)", "I+D"),
        ]),
        ("🔄 PROTOCOLO DE PRUEBA (3 iteraciones mín.)", [
            ("Iteración 1: ejecutar receta, fotografiar resultado", "I+D"),
            ("Iteración 1: feedback equipo cocina (sabor, textura, visual)", "I+D"),
            ("Iteración 1: registrar ajustes necesarios", "I+D"),
            ("Iteración 2: aplicar ajustes, fotografiar resultado", "I+D"),
            ("Iteración 2: feedback equipo cocina + sala", "I+D"),
            ("Iteración 2: registrar mejoras y validar concepto", "I+D"),
            ("Iteración 3: versión final, foto emplatado definitivo", "I+D"),
            ("Iteración 3: prueba de tiempos en simulación de servicio", "I+D"),
            ("Iteración 3: aprobación del chef ejecutivo", "I+D"),
        ]),
    ])
    create_task_sheet(wb, "Documentación", "3F51B5", "Documentación de Receta Final", [
        ("📋 FICHA TÉCNICA COMPLETA", [
            ("Nombre definitivo del plato", "Fichas"),
            ("Ingredientes exactos con gramajes (0,01g para hidrocoloides)", "Fichas"),
            ("Proceso paso a paso numerado", "Fichas"),
            ("Temperaturas y tiempos exactos de cada fase", "Fichas"),
            ("Foto de emplatado final (ángulo cenital + 45°)", "Fichas"),
            ("Foto de corte/interior (si aplica)", "Fichas"),
            ("Alérgenos presentes en el plato final", "Fichas"),
            ("Food cost real por ración", "Food cost"),
            ("Margen objetivo cumplido (sí/no)", "Food cost"),
        ]),
        ("⚙️ VIABILIDAD OPERATIVA", [
            ("¿Ejecutable en servicio de 40 covers? (sí/no, ajustes)", "I+D"),
            ("Tiempo total de emplatado por ración", "Emplatado"),
            ("Número de cocineros necesarios para este pase", "I+D"),
            ("Pre-elaboraciones requeridas y tiempo total", "I+D"),
            ("Equipamiento especial necesario (disponible sí/no)", "I+D"),
            ("Proveedor del ingrediente principal confirmado", "Proveedores"),
            ("Disponibilidad estacional del ingrediente clave", "Proveedores"),
        ]),
        ("📅 CALENDARIO ESTACIONAL (4 cambios/año)", [
            ("Primavera (mar-may): definir platos nuevos y retirar", "I+D"),
            ("Verano (jun-ago): definir platos nuevos y retirar", "I+D"),
            ("Otoño (sep-nov): definir platos nuevos y retirar", "I+D"),
            ("Invierno (dic-feb): definir platos nuevos y retirar", "I+D"),
            ("Registrar proveedores por ingrediente estacional", "Proveedores"),
            ("Actualizar plating bible con cada cambio de carta", "Fichas"),
        ]),
        ("📚 BIBLIOTECA DE TÉCNICAS", [
            ("Registrar técnica probada: nombre, resultado, aplicación", "Técnica"),
            ("Registrar técnica descartada: razón del descarte", "Técnica"),
            ("Actualizar catálogo de hidrocoloides probados y ratios", "Técnica"),
            ("Actualizar catálogo de fermentaciones activas y tiempos", "Técnica"),
            ("Documentar maridajes técnica-ingrediente exitosos", "Técnica"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "03-id-desarrollo-menu.xlsx"))
    print("✅ 03-id-desarrollo-menu.xlsx")


# ═══════════════════════════════════════════════════════════
# 04 — TAREAS POR ROL — BRIGADA CREATIVA
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Tareas por Rol — Brigada Creativa", [
        "Cómo usar esta plantilla:",
        "▸ Asigna responsabilidades claras a cada miembro de la brigada.",
        "▸ Cada rol tiene tareas diarias específicas pre-servicio y servicio.",
        "▸ El chef ejecutivo valida el pase final.",
        "▸ La coordinación cocina-sala es CRÍTICA en menú degustación.",
        "",
        "Roles clave:",
        "▸ Chef Ejecutivo — visión creativa, I+D, pase final",
        "▸ Sous-Chef — gestión operativa, mise en place",
        "▸ Chef Pastelero — postres de autor, helados, fermentaciones dulces",
        "▸ Sumiller — maridaje, bodega, formación sala",
        "▸ Maître — storytelling, tiempos de mesa, coordinación",
    ])
    create_task_sheet(wb, "Chef Ejecutivo", "FF9800", "Tareas Diarias — Chef Ejecutivo", [
        ("👨‍🍳 CHEF EJECUTIVO — DIARIO", [
            ("Revisión del menú del día: ajustes, sustituciones estacionales", "Cocina"),
            ("Sesión I+D programada (si es día asignado)", "I+D"),
            ("Contacto con proveedores premium (trufa, marisco, producto vivo)", "Proveedores"),
            ("Revisión de food cost semanal", "Food cost"),
            ("Aprobación de mise en place antes del servicio", "Cocina"),
            ("Supervisión de emplatado: verificar plating bible", "Emplatado"),
            ("Pase final: control de calidad de cada plato antes de sala", "Pase"),
            ("Comunicación con maître sobre tiempos y secuencia", "Maître"),
            ("Feedback post-servicio al equipo", "Cocina"),
        ]),
    ])
    create_task_sheet(wb, "Sous-Chef", "FF9800", "Tareas Diarias — Sous-Chef", [
        ("🔥 SOUS-CHEF — DIARIO", [
            ("Coordinación de mise en place de todas las partidas", "Mise en place"),
            ("Gestión de la brigada: asignación de tareas por turno", "Cocina"),
            ("Supervisión de pre-elaboraciones largo recorrido", "Cocina"),
            ("Control de temperaturas (sous-vide, fermentaciones, cámaras)", "Cámara"),
            ("Verificar etiquetado de TODAS las pre-elaboraciones", "Cocina"),
            ("Reposición de stock durante servicio", "Stock"),
            ("Apoyo en emplatado durante pico de servicio", "Emplatado"),
            ("Revisión de pedidos a proveedores para mañana", "Proveedores"),
            ("Cierre de cocina: checklist de equipamiento y limpieza", "Limpieza"),
        ]),
    ])
    create_task_sheet(wb, "Partida Caliente", "FF9800", "Tareas Diarias — Chef de Partida Caliente", [
        ("🔥 PARTIDA CALIENTE — DIARIO", [
            ("Preparación de fondos y salsas del día", "Caliente"),
            ("Cocciones principales: proteínas, guarniciones calientes", "Caliente"),
            ("Gestión de sous-vide en servicio (regenerar, sellar)", "Caliente"),
            ("Acabados de carne: sellado, glaseado, reposo", "Caliente"),
            ("Acabados de pescado: a la plancha, brasa, vapor", "Caliente"),
            ("Salsas calientes al momento (reducir, montar)", "Caliente"),
            ("Coordinación de tiempos con pase y emplatado", "Pase"),
        ]),
    ])
    create_task_sheet(wb, "Partida Fría", "2196F3", "Tareas Diarias — Chef de Partida Fría / Creativo", [
        ("❄️ PARTIDA FRÍA / CREATIVO — DIARIO", [
            ("Ensaladas conceptuales: montaje, aliño al momento", "Frío"),
            ("Ceviches y tartares: cortar máx. 2h antes de servicio", "Frío"),
            ("Crudos y carpaccios: laminar, temperar", "Frío"),
            ("Marinados rápidos (cítricos, vinagres)", "Frío"),
            ("Emulsiones frías, vinagretas de autor", "Frío"),
            ("Montaje de pases fríos según plating bible", "Emplatado"),
            ("Control de temperatura de producto en línea fría", "Cámara"),
        ]),
    ])
    create_task_sheet(wb, "Chef Pastelero", "E91E63", "Tareas Diarias — Chef Pastelero Creativo", [
        ("🍰 CHEF PASTELERO CREATIVO — DIARIO", [
            ("Pre-postres: preparación y montaje de componentes", "Pastelería"),
            ("Postre principal: mise en place y emplatado", "Pastelería"),
            ("Petit fours: producción y montaje para servicio", "Petit fours"),
            ("Helados de autor: turbinar, almacenar, temperar", "Pastelería"),
            ("Sorbetes y nieves: producción del día", "Pastelería"),
            ("Fermentaciones dulces: verificar estado (kombucha, kéfir)", "I+D"),
            ("Crujientes y texturas dulces: tuiles, snaps, caramelos", "Pastelería"),
            ("Verificar plating bible de postres: foto y posición", "Emplatado"),
            ("Coordinación con cocina para timing pre-postre/postre", "Pase"),
        ]),
    ])
    create_task_sheet(wb, "Sumiller", "9C27B0", "Tareas Diarias — Sumiller", [
        ("🍷 SUMILLER — DIARIO", [
            ("Verificar maridaje del día: ajustar si hay cambios en menú", "Maridaje"),
            ("Preparar vinos del maridaje: temperatura, decantación", "Bodega"),
            ("Preparar maridaje alternativo (mocktails, kombucha, cerveza)", "Maridaje"),
            ("Briefing a equipo de sala sobre maridaje y storytelling", "Sala"),
            ("Servicio de vinos sincronizado con cocina", "Servicio"),
            ("Control de stock de bodega post-servicio", "Bodega"),
            ("Registrar consumo de vinos del día", "Bodega"),
        ]),
    ])
    create_task_sheet(wb, "Maître-Sala", "2196F3", "Tareas Diarias — Maître y Equipo de Sala", [
        ("🎩 MAÎTRE / JEFE DE SALA — DIARIO", [
            ("Storytelling: preparar narrativa de cada pase para comensales", "Storytelling"),
            ("Tiempos de mesa: coordinar con cocina secuencia de pases", "Maître"),
            ("Gestión de VIPs: atención especial, preferencias", "Maître"),
            ("Supervisión del servicio sincronizado (mesa ↔ cocina)", "Servicio"),
            ("Gestión de incidencias en sala (alérgenos, quejas, peticiones)", "Maître"),
            ("Despedida personalizada a cada mesa", "Sala"),
        ]),
        ("🍽️ CAMARERO MENÚ DEGUSTACIÓN — DIARIO", [
            ("Servicio sincronizado: recibir pase de cocina y servir", "Camarero"),
            ("Explicación de cada pase al comensal (storytelling breve)", "Storytelling"),
            ("Verificar alérgenos por mesa ANTES de cada pase", "Camarero"),
            ("Retirada de platos coordinada con cocina", "Camarero"),
            ("Servicio de pan, agua, extras entre pases", "Camarero"),
            ("Comunicar feedback de comensales al maître", "Camarero"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "04-tareas-brigada-creativa.xlsx"))
    print("✅ 04-tareas-brigada-creativa.xlsx")


# ═══════════════════════════════════════════════════════════
# 05 — TAREAS SEMANALES Y MENSUALES
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Tareas Semanales y Mensuales", [
        "Cómo usar esta plantilla:",
        "▸ Programa las tareas semanales en un día fijo (lunes recomendado).",
        "▸ Las tareas mensuales se programan el primer lunes del mes.",
        "▸ Registra cada revisión con fecha y firma del responsable.",
        "",
        "Puntos clave:",
        "▸ Inventario semanal de productos premium (trufa, caviar, etc.).",
        "▸ Mantenimiento preventivo de equipamiento de vanguardia.",
        "▸ Revisión de food cost real vs. objetivo cada mes.",
    ])
    create_task_sheet(wb, "Semanales", "4CAF50", "Tareas Semanales", [
        ("📦 INVENTARIO Y STOCK SEMANAL", [
            ("Inventario productos premium (trufa, caviar, azafrán, vainilla)", "Stock"),
            ("Inventario hidrocoloides y texturizantes (agar, gellan, xantana)", "Stock"),
            ("Inventario gases (N2O sifones, CO2, butano sopletes)", "Stock"),
            ("Verificar stock de microbrotes y flores comestibles", "Stock"),
            ("Pedido semanal a proveedores premium", "Proveedores"),
            ("Registrar mermas de la semana y causas", "Admin"),
        ]),
        ("🧪 FERMENTACIONES Y MADURACIONES", [
            ("Medir pH de TODAS las fermentaciones activas", "I+D"),
            ("Verificar estado de koji / cultivos activos", "I+D"),
            ("Controlar maduraciones en cámara (temp., humedad, día)", "Cámara"),
            ("Evaluar fermentaciones listas para uso en carta", "I+D"),
            ("Descartar fermentaciones fuera de parámetros", "I+D"),
        ]),
        ("🔧 MANTENIMIENTO EQUIPAMIENTO", [
            ("Mantenimiento Pacojet: limpieza profunda, verificar cuchillas", "Cocina"),
            ("Mantenimiento Roner/sous-vide: calibración, limpieza", "Cocina"),
            ("Mantenimiento deshidratador: limpieza bandejas, filtros", "Cocina"),
            ("Verificar calibración de básculas de precisión (0,01g)", "Cocina"),
            ("Mantenimiento sifones: juntas, válvulas, limpieza", "Cocina"),
        ]),
        ("👥 EQUIPO Y FORMACIÓN", [
            ("Reunión semanal de equipo (feedback, planificación)", "Cocina"),
            ("Revisión de fichas técnicas actualizadas", "Fichas"),
            ("Actualización de plating bible si hubo cambios", "Emplatado"),
        ]),
    ])
    create_task_sheet(wb, "Mensuales", "4CAF50", "Tareas Mensuales", [
        ("🔄 CARTA Y MENÚ", [
            ("Rotación parcial carta: evaluar 2-3 platos para cambio", "I+D"),
            ("Evaluación de platos con menor aceptación", "I+D"),
            ("Planificación de nuevos platos para próximo mes", "I+D"),
            ("Actualización de plating bible completa", "Fichas"),
        ]),
        ("🧹 DEEP CLEAN Y CALIBRACIÓN", [
            ("Deep clean del laboratorio / zona I+D", "Limpieza"),
            ("Deep clean de cámaras de maduración", "Limpieza"),
            ("Calibración profesional de básculas y termómetros", "Cocina"),
            ("Revisión de todos los equipos de vanguardia", "Cocina"),
            ("Mantenimiento de cámaras frigoríficas (filtros, juntas)", "Limpieza"),
        ]),
        ("💰 FINANZAS Y GESTIÓN", [
            ("Inventario completo de bodega", "Bodega"),
            ("Revisión de proveedores: calidad, precio, entregas", "Proveedores"),
            ("Food cost real vs. objetivo (análisis por plato)", "Food cost"),
            ("Coste de bebida por cubierto", "Food cost"),
            ("Revisión de mermas mensuales y plan de reducción", "Admin"),
        ]),
        ("📚 FORMACIÓN Y DESARROLLO", [
            ("Sesión de formación equipo (técnica nueva o repaso)", "I+D"),
            ("Evaluación de desempeño informal por rol", "Admin"),
            ("Planificación de eventos especiales del próximo mes", "Eventos"),
            ("Revisión de objetivos y KPIs del mes", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx"))
    print("✅ 05-tareas-semanales-mensuales.xlsx")


# ═══════════════════════════════════════════════════════════
# 06 — SUMILLER Y MARIDAJES
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Sumiller y Maridajes", [
        "Cómo usar esta plantilla:",
        "▸ Diseña el maridaje para cada pase del menú degustación.",
        "▸ Gestiona la bodega con control de stock y temperaturas.",
        "▸ Incluye opciones de maridaje alternativo (sin alcohol).",
        "▸ Forma al equipo de sala en el storytelling de cada vino.",
        "",
        "Protocolo de servicio:",
        "▸ Copa correcta por tipo de vino/bebida.",
        "▸ Temperatura exacta de servicio.",
        "▸ Cantidad por copa en maridaje (≈60-80ml por pase).",
    ])
    create_task_sheet(wb, "Maridaje Menú", "9C27B0", "Diseño de Maridaje por Pase", [
        ("🍷 MARIDAJE — MENÚ DEGUSTACIÓN", [
            ("Pase 1 — Snack: vino/bebida seleccionada, copa, temp., ml", "Maridaje"),
            ("Pase 2 — Aperitivo: vino/bebida, copa, temp., ml", "Maridaje"),
            ("Pase 3 — Entrante frío: vino/bebida, copa, temp., ml", "Maridaje"),
            ("Pase 4 — Entrante caliente: vino/bebida, copa, temp., ml", "Maridaje"),
            ("Pase 5 — Pescado: vino/bebida, copa, temp., ml", "Maridaje"),
            ("Pase 6 — Carne: vino/bebida, copa, temp., ml", "Maridaje"),
            ("Pase 7 — Pre-postre: vino/bebida, copa, temp., ml", "Maridaje"),
            ("Pase 8 — Postre: vino/bebida, copa, temp., ml", "Maridaje"),
        ]),
        ("🍹 MARIDAJE ALTERNATIVO (sin alcohol)", [
            ("Pase 1 — Snack: mocktail/kombucha/zumo, descripción", "Maridaje"),
            ("Pase 2 — Aperitivo: alternativa sin alcohol", "Maridaje"),
            ("Pase 3-4 — Entrantes: alternativa sin alcohol", "Maridaje"),
            ("Pase 5 — Pescado: alternativa sin alcohol", "Maridaje"),
            ("Pase 6 — Carne: alternativa sin alcohol", "Maridaje"),
            ("Pase 7-8 — Postres: alternativa sin alcohol", "Maridaje"),
            ("Cócteles de autor (con alcohol): lista y recetas", "Maridaje"),
            ("Cervezas artesanas seleccionadas por pase (si aplica)", "Maridaje"),
        ]),
    ])
    create_task_sheet(wb, "Gestión Bodega", "9C27B0", "Gestión de Bodega", [
        ("🏪 STOCK Y ROTACIÓN", [
            ("Inventario completo de bodega (actualizar mensual)", "Bodega"),
            ("Control de stock mínimo por referencia del maridaje", "Bodega"),
            ("Rotación FIFO estricta (primero en entrar, primero en salir)", "Bodega"),
            ("Temperatura de bodega verificada (zona tintos, blancos, espumosos)", "Bodega"),
            ("Humedad de bodega verificada", "Bodega"),
            ("Registrar entradas de vino (fecha, proveedor, precio)", "Bodega"),
            ("Registrar salidas de vino (fecha, servicio, cantidad)", "Bodega"),
        ]),
        ("📋 FICHAS DE CATA", [
            ("Ficha de cata por cada vino del maridaje activo", "Sumiller"),
            ("Notas de cata para storytelling de sala", "Sumiller"),
            ("Temperatura óptima de servicio registrada", "Sumiller"),
            ("Tiempo de decantación (si aplica)", "Sumiller"),
            ("Maridaje recomendado (plato + técnica)", "Sumiller"),
        ]),
        ("🤝 PROVEEDORES Y FORMACIÓN", [
            ("Relación con bodegas: contacto, condiciones, rappel", "Proveedores"),
            ("Relación con distribuidores: catálogo, entregas, crédito", "Proveedores"),
            ("Formación semanal equipo de sala (vino del maridaje)", "Sala"),
            ("Control coste de bebida por cubierto", "Food cost"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "06-sumiller-maridajes.xlsx"))
    print("✅ 06-sumiller-maridajes.xlsx")


# ═══════════════════════════════════════════════════════════
# 07 — CHEF'S TABLE Y EVENTOS ESPECIALES
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Chef's Table y Eventos Especiales", [
        "Cómo usar esta plantilla:",
        "▸ Organiza cada tipo de evento con su checklist específico.",
        "▸ El chef's table requiere coordinación directa chef-comensales.",
        "▸ Los pop-ups necesitan logística especial (equipamiento portátil).",
        "▸ Planifica el calendario anual de eventos desde enero.",
        "",
        "Tipos de evento:",
        "▸ Chef's table — experiencia premium con interacción directa",
        "▸ Pop-ups — logística portátil, mise en place adaptada",
        "▸ Cenas maridaje — colaboración con bodegas invitadas",
        "▸ Showcookings — demostraciones para público",
        "▸ Cenas privadas — menú, decoración, facturación",
    ])
    create_task_sheet(wb, "Chef's Table", "00BCD4", "Checklist Chef's Table", [
        ("🍽️ CHEF'S TABLE — PREPARACIÓN", [
            ("Confirmar número de comensales y alérgenos", "Chef's table"),
            ("Diseñar menú especial chef's table (ampliado o exclusivo)", "Cocina"),
            ("Preparar storytelling de cada pase (chef directo)", "Storytelling"),
            ("Mise en place visible: organizar estación para interacción", "Mise en place"),
            ("Verificar vajilla especial / piezas exclusivas", "Sala"),
            ("Coordinar con sumiller maridaje especial", "Maridaje"),
        ]),
        ("🎭 CHEF'S TABLE — SERVICIO", [
            ("Presentación del chef a los comensales", "Chef's table"),
            ("Interacción durante emplatado (explicar técnica en vivo)", "Chef's table"),
            ("Timing de pases adaptado al ritmo de la mesa", "Chef's table"),
            ("Ofrecer bocados extra / fuera de carta (sorpresa del chef)", "Chef's table"),
            ("Fotografía de la experiencia (con permiso del comensal)", "Fotografía"),
            ("Despedida personalizada del chef", "Chef's table"),
        ]),
    ])
    create_task_sheet(wb, "Pop-ups", "00BCD4", "Checklist Pop-ups y Eventos Externos", [
        ("🚐 POP-UP — LOGÍSTICA", [
            ("Seleccionar ubicación y verificar instalaciones (agua, luz, gas)", "Eventos"),
            ("Lista de equipamiento portátil necesario", "Eventos"),
            ("Transporte de equipamiento y producto", "Eventos"),
            ("Mise en place adaptada a las limitaciones del espacio", "Mise en place"),
            ("Menú adaptado al formato pop-up (simplificado pero impactante)", "Cocina"),
            ("Personal necesario: cocina + sala + logística", "Eventos"),
        ]),
        ("🎤 SHOWCOOKINGS Y DEMOSTRACIONES", [
            ("Definir tema/técnica de la demostración", "Eventos"),
            ("Preparar guion y timing (45-90 min habitual)", "Eventos"),
            ("Lista de ingredientes y equipamiento necesario", "Eventos"),
            ("Verificar instalaciones del lugar (corriente, agua, extractor)", "Eventos"),
            ("Material de apoyo (pantalla, cámara cenital, micrófono)", "Eventos"),
        ]),
    ])
    create_task_sheet(wb, "Cenas Especiales", "00BCD4", "Cenas Maridaje y Cenas Privadas", [
        ("🍷 CENAS MARIDAJE CON BODEGA INVITADA", [
            ("Contactar bodega: seleccionar vinos, condiciones", "Sumiller"),
            ("Diseñar menú que potencie cada vino del maridaje", "Cocina"),
            ("Definir precio del evento y número de plazas", "Admin"),
            ("Material de comunicación (RRSS, newsletter, web)", "RRSS"),
            ("Preparar fichas de cata para comensales", "Sumiller"),
            ("Coordinar presencia del enólogo/bodeguero en sala", "Eventos"),
        ]),
        ("🕯️ CENAS PRIVADAS", [
            ("Menú personalizado según solicitud del cliente", "Cocina"),
            ("Alérgenos y preferencias confirmados", "Cocina"),
            ("Servicio exclusivo: personal asignado", "Servicio"),
            ("Decoración especial de la mesa/sala", "Sala"),
            ("Presupuesto aprobado y facturación preparada", "Admin"),
        ]),
        ("📅 CALENDARIO ANUAL DE EVENTOS", [
            ("Planificar eventos por trimestre (mín. 1/mes)", "Eventos"),
            ("Presupuesto anual de eventos especiales", "Admin"),
            ("Calendario de colaboraciones con bodegas", "Sumiller"),
            ("Calendario de pop-ups y showcookings", "Eventos"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "07-chefs-table-eventos.xlsx"))
    print("✅ 07-chefs-table-eventos.xlsx")


# ═══════════════════════════════════════════════════════════
# 08 — FOTOGRAFÍA Y STORYTELLING
# ═══════════════════════════════════════════════════════════
def gen_08():
    wb = Workbook()
    add_instructions_sheet(wb, "08 · Fotografía, RRSS y Storytelling", [
        "Cómo usar esta plantilla:",
        "▸ Fotografía cada plato nuevo ANTES de incluirlo en carta.",
        "▸ Mantén un calendario editorial de RRSS constante.",
        "▸ El storytelling es parte del producto: documéntalo.",
        "▸ Gestiona la relación con prensa y guías gastronómicas.",
        "",
        "Frecuencia recomendada:",
        "▸ 4-7 posts/semana en feed (Instagram/TikTok).",
        "▸ 2-3 Stories diarias (behind the scenes).",
        "▸ 1 Reel/TikTok semanal (proceso creativo).",
    ])
    create_task_sheet(wb, "Fotografía Platos", "8BC34A", "Sesión de Fotografía de Platos", [
        ("📸 SESIÓN FOTOGRÁFICA — CADA PLATO NUEVO", [
            ("Preparar emplatado perfecto según plating bible", "Emplatado"),
            ("Foto cenital (90°) con iluminación controlada", "Fotografía"),
            ("Foto ángulo 45° (la más usada en RRSS)", "Fotografía"),
            ("Foto detalle / macro (textura, corte, acabado)", "Fotografía"),
            ("Foto de corte / interior (si aplica)", "Fotografía"),
            ("Verificar fondos y props (plato de autor, mantel, cubiertos)", "Fotografía"),
            ("Edición y retoque (brillo, contraste, color consistente)", "Fotografía"),
            ("Guardar en carpeta organizada por plato y temporada", "Fotografía"),
        ]),
        ("📱 KIT FOTOGRAFÍA — ESTÁNDARES", [
            ("Ángulos definidos para el restaurante (cenital, 45°, lateral)", "Fotografía"),
            ("Iluminación: natural o softbox, dirección consistente", "Fotografía"),
            ("Fondos: madera oscura, piedra, tela según estilo del restaurante", "Fotografía"),
            ("Props: vajilla de autor, cubiertos, cristalería, vegetales", "Fotografía"),
        ]),
    ])
    create_task_sheet(wb, "RRSS", "8BC34A", "Calendario Editorial y RRSS", [
        ("📅 CALENDARIO EDITORIAL SEMANAL", [
            ("Lunes: post de plato de la carta (foto profesional)", "RRSS"),
            ("Martes: Story behind the scenes (cocina, I+D)", "RRSS"),
            ("Miércoles: Reel / TikTok (proceso creativo, emplatado)", "RRSS"),
            ("Jueves: Story del equipo / proveedor", "RRSS"),
            ("Viernes: post de preparación del fin de semana", "RRSS"),
            ("Sábado: Stories de servicio en directo", "RRSS"),
            ("Domingo: post de inspiración / filosofía del chef", "RRSS"),
        ]),
        ("🎬 CONTENIDO BEHIND-THE-SCENES", [
            ("Sesión I+D filmada (proceso de creación de plato)", "RRSS"),
            ("Emplatado en detalle (cámara cenital de pase)", "RRSS"),
            ("Visita a proveedor (mercado, huerta, pescadería)", "RRSS"),
            ("Equipo: presentar a cada miembro de la brigada", "RRSS"),
            ("Técnica explicada (sous-vide, fermentación, esferificación)", "RRSS"),
        ]),
    ])
    create_task_sheet(wb, "Prensa-Guías", "8BC34A", "Prensa y Guías Gastronómicas", [
        ("📰 NARRATIVA Y STORYTELLING", [
            ("Narrativa de cada plato documentada (para carta, web, sala)", "Storytelling"),
            ("Historia del restaurante actualizada (para prensa)", "Marketing"),
            ("Biografía del chef actualizada", "Marketing"),
            ("Filosofía culinaria en 3-5 frases (para comunicación)", "Marketing"),
        ]),
        ("⭐ GESTIÓN DE RESEÑAS Y PRENSA", [
            ("Monitorizar reseñas Google, TripAdvisor, TheFork", "RRSS"),
            ("Responder reseñas (positivas y negativas) en 24-48h", "RRSS"),
            ("Gestión de prensa gastronómica (notas, invitaciones)", "Prensa"),
            ("Material para guías: Michelin, Repsol, 50 Best, OAD", "Prensa"),
            ("Dossier de prensa actualizado (fotos, textos, hitos)", "Prensa"),
            ("Gestión de influencers y food bloggers (invitaciones)", "RRSS"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "08-fotografia-storytelling.xlsx"))
    print("✅ 08-fotografia-storytelling.xlsx")


# ═══════════════════════════════════════════════════════════
# 09 — PLANTILLA PERSONALIZABLE (3 blank templates)
# ═══════════════════════════════════════════════════════════
def gen_09():
    wb = Workbook()
    add_instructions_sheet(wb, "09 · Plantilla Personalizable", [
        "Cómo usar esta plantilla:",
        "▸ Usa estas hojas en blanco para crear tus propios checklists.",
        "▸ Copia la estructura de cualquier otro archivo del kit.",
        "▸ Añade tareas específicas de tu restaurante creativo.",
        "▸ Las 3 plantillas son idénticas para que tengas margen.",
        "",
        "Ideas de uso:",
        "▸ Checklist de cambio de carta estacional",
        "▸ Protocolo de visita de inspección o guía",
        "▸ Tareas específicas de un evento puntual",
    ])
    col_headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
    create_blank_template_sheet(wb, "Plantilla A", "FFD700", "Plantilla Personalizable A", col_headers)
    create_blank_template_sheet(wb, "Plantilla B", "FFD700", "Plantilla Personalizable B", col_headers)
    create_blank_template_sheet(wb, "Plantilla C", "FFD700", "Plantilla Personalizable C", col_headers)
    wb.save(os.path.join(OUTPUT_DIR, "09-plantilla-personalizable.xlsx"))
    print("✅ 09-plantilla-personalizable.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS-01 — BRIEFING DIARIO DE SERVICIO
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS · Briefing Diario de Servicio", [
        "Cómo usar esta plantilla:",
        "▸ Rellena ANTES de cada servicio (almuerzo y/o cena).",
        "▸ El chef ejecutivo y el maître revisan juntos.",
        "▸ Se comparte con TODO el equipo (cocina + sala).",
        "▸ Es el documento central de comunicación del servicio.",
        "",
        "Elementos clave del briefing:",
        "▸ Menú del día con posibles cambios o sustituciones.",
        "▸ Alérgenos de reservas confirmadas.",
        "▸ VIPs, prensa, guías, eventos especiales.",
        "▸ Storytelling actualizado de cada pase.",
    ])
    create_task_sheet(wb, "Briefing Servicio", "FF5722", "Briefing Diario Pre-Servicio", [
        ("📋 MENÚ Y CAMBIOS DEL DÍA", [
            ("Menú degustación del día: confirmar todos los pases", "Cocina"),
            ("Cambios o sustituciones respecto al día anterior", "Cocina"),
            ("Platos en prueba (I+D) que se sirven hoy (si aplica)", "I+D"),
            ("Disponibilidad de producto: confirmado todo el stock", "Stock"),
            ("Platos con producción limitada (indicar máximo de raciones)", "Cocina"),
        ]),
        ("⚠️ ALÉRGENOS Y RESTRICCIONES", [
            ("Revisar reservas: listar alérgenos por mesa", "Maître"),
            ("Identificar mesas con restricciones especiales (vegano, halal, etc.)", "Maître"),
            ("Preparar alternativas para mesas con alérgenos", "Cocina"),
            ("Comunicar al equipo de sala cada restricción por mesa", "Sala"),
        ]),
        ("⭐ VIPs, PRENSA Y EVENTOS", [
            ("Identificar VIPs del día (nombre, mesa, preferencias)", "Maître"),
            ("Prensa/guías/influencers: protocolo de atención", "Maître"),
            ("Chef's table hoy: confirmar menú, timing, storytelling", "Chef's table"),
            ("Cena privada / evento especial: detalles confirmados", "Eventos"),
        ]),
        ("🍷 MARIDAJE Y STORYTELLING", [
            ("Maridaje del día confirmado por el sumiller", "Maridaje"),
            ("Cambios en maridaje respecto al día anterior", "Sumiller"),
            ("Storytelling de cada pase: repaso con equipo de sala", "Storytelling"),
            ("Datos del vino/bodega del día para conversación con comensal", "Sumiller"),
        ]),
        ("🧪 PLATOS EN PRUEBA / I+D", [
            ("Plato en prueba hoy (si se ofrece a mesa seleccionada)", "I+D"),
            ("Feedback a recoger: sabor, textura, presentación, opinión", "I+D"),
            ("Responsable de recoger feedback de sala", "Sala"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))
    print("✅ BONUS-01-briefing-servicio.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS-02 — CALENDARIO ANUAL
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS · Calendario Anual — Restaurante Creativo", [
        "Cómo usar esta plantilla:",
        "▸ Planifica el año completo con cambios estacionales de carta.",
        "▸ Identifica fechas clave para menús especiales.",
        "▸ Coordina participación en festivales gastronómicos.",
        "▸ Anticipa la preparación con timeline suficiente.",
        "",
        "Cambios de carta recomendados:",
        "▸ 4 cambios estacionales principales.",
        "▸ Rotación parcial mensual (2-3 platos).",
        "▸ Menús especiales para fechas señaladas.",
    ])
    create_calendar_sheet(wb, "Calendario Anual", "FF5722", "Calendario Anual — Restaurante Creativo / De Autor", [
        ("Enero", "Carta de invierno consolidada. Trufa negra (temporada alta). Caza menor. Menú especial Año Nuevo (1 ene). Planificación anual de eventos.", "Trufa negra, caza, tubérculos", "Media"),
        ("Febrero", "San Valentín: menú especial degustación parejas. Trufa negra (última fase). Cítricos de temporada. Prep. cambio carta primavera.", "Trufa, cítricos, alcachofas", "Alta"),
        ("Marzo", "CAMBIO CARTA PRIMAVERA. Espárragos, guisantes, habas. Sesiones I+D nuevos platos primavera. Festival gastronómico local (revisar).", "Espárragos, guisantes, cordero", "Media"),
        ("Abril", "Carta primavera rodada. Semana Santa: aforo alto, menú especial. Alcachofas, flores comestibles. Preparar pop-up de temporada.", "Alcachofas, flores, fresas", "Alta"),
        ("Mayo", "Cerezas, fresas, nísperos. Espárragos trigueros. Atún rojo (inicio temporada). Preparar cambio carta verano.", "Atún rojo, cerezas, espárragos", "Alta"),
        ("Junio", "CAMBIO CARTA VERANO. Tomates de temporada, melocotón, albaricoque. Gazpachos y sopas frías de autor. Terraza / cenas al aire libre.", "Tomate, melocotón, gambas", "Alta"),
        ("Julio", "Carta de verano consolidada. Higos, melón, sandía. Maridaje con vinos frescos y rosados. Turismo gastronómico: aforo alto.", "Higos, sandía, pescados azules", "Alta"),
        ("Agosto", "Verano alta ocupación. Pimientos, berenjenas, calabacín. Helados de autor (producción extra). Prep. cambio carta otoño.", "Pimiento, berenjena, helados", "Alta"),
        ("Septiembre", "CAMBIO CARTA OTOÑO. Setas silvestres (inicio temporada). Uvas, higos tardíos. Cenas maridaje vendimia con bodegas.", "Setas, uvas, higos", "Alta"),
        ("Octubre", "Temporada plena de setas (boletus, rebozuelo, trompeta). Calabaza, boniato. Castañas. Festival gastronómico otoño.", "Setas, calabaza, castañas", "Alta"),
        ("Noviembre", "Trufa negra (inicio temporada). Caza mayor y menor. Prep. menús Navidad. Reservas navideñas abiertas.", "Trufa, caza, raíces", "Media"),
        ("Diciembre", "CAMBIO CARTA INVIERNO. Menús Navidad, Nochebuena, Nochevieja. Trufa, foie, marisco, caviar. Máxima facturación del año.", "Trufa, foie, marisco, caviar", "Alta"),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual.xlsx"))
    print("✅ BONUS-02-calendario-anual.xlsx")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("🚀 Generando Kit de Tareas: Restaurante Creativo / De Autor...")
    print(f"📁 Output: {OUTPUT_DIR}\n")
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_08()
    gen_09()
    gen_bonus_01()
    gen_bonus_02()
    print(f"\n🎉 ¡11 archivos generados correctamente!")
    print(f"📂 Ruta: {OUTPUT_DIR}")
