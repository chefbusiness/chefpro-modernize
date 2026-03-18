#!/usr/bin/env python3
"""
Generate APPCC (HACCP) compliance templates for Pack de Plantillas APPCC.
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Output ──
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "pack-appcc"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Brand Colors ──
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
RED_ALERT = "FF4444"
GREEN_OK = "4CAF50"
ORANGE_WARN = "FF9800"
LIGHT_GREEN = "E8F5E9"
LIGHT_RED = "FFEBEE"
LIGHT_BLUE = "E3F2FD"
LIGHT_YELLOW = "FFF8E1"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# ── Fonts ──
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
alert_font = Font(name="Calibri", size=11, bold=True, color=RED_ALERT)

# ── Fills ──
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
input_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
light_row = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
alt_row = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")

# ── Borders ──
thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

# ── Alignments ──
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")


def add_instructions_sheet(wb, title, lines):
    """Add a branded instructions sheet as the first tab."""
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80

    ws.merge_cells("B2:B2")
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)

    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")

    return ws


def style_header_row(ws, row, cols, headers):
    """Apply header styling to a row."""
    for col_idx, h in enumerate(headers, 1):
        if col_idx > cols:
            break
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def add_data_rows(ws, start_row, num_rows, num_cols, input_cols=None):
    """Add empty styled data rows."""
    if input_cols is None:
        input_cols = list(range(1, num_cols + 1))
    for i in range(num_rows):
        row = start_row + i
        fill = light_row if i % 2 == 0 else alt_row
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = data_font
            cell.alignment = center_align
            if col in input_cols:
                cell.fill = input_fill


def add_footer(ws, row, text):
    """Add a footer note."""
    ws.cell(row=row, column=1, value=text).font = small_font


# ══════════════════════════════════════════════════════════════
# TEMPLATE 1: Registro de Temperaturas Diario
# ══════════════════════════════════════════════════════════════
def create_temperaturas_diario():
    wb = Workbook()
    add_instructions_sheet(wb, "Registro de Temperaturas — Control Diario", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra la temperatura de cada equipo 2 veces al día (mañana y tarde)",
        "▸ Las celdas verdes son editables — introduce la temperatura leída",
        "▸ La columna 'Estado' se calcula automáticamente (OK / ALERTA)",
        "▸ Si la temperatura sale del rango, aparece ALERTA en rojo",
        "▸ Imprime la hoja semanal y archívala firmada",
        "",
        "Rangos de temperatura obligatorios",
        "",
        "▸ Cámara frigorífica: 0°C a 4°C",
        "▸ Congelador: -18°C o inferior",
        "▸ Exposición caliente: 65°C o superior",
        "▸ Exposición fría: 0°C a 8°C",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    # Weekly sheet for each equipment type
    equipos = [
        ("Cámara 1 (Refrigeración)", 0, 4),
        ("Cámara 2 (Refrigeración)", 0, 4),
        ("Congelador 1", -25, -18),
        ("Congelador 2", -25, -18),
        ("Exposición Fría (Vitrina)", 0, 8),
        ("Exposición Caliente (Baño María)", 65, 100),
    ]

    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    ws = wb.create_sheet(title="Registro Semanal")
    ws.sheet_properties.tabColor = "2196F3"

    ws.column_dimensions["A"].width = 30
    for c in range(2, 22):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Registro de Temperaturas — Control Diario"
    ws["A1"].font = title_font

    ws["A2"].value = "Semana del: ___/___/______ al ___/___/______"
    ws["A2"].font = subtitle_font

    ws["A3"].value = "Responsable: _________________________________"
    ws["A3"].font = subtitle_font

    # For each equipment
    current_row = 5
    for equipo, temp_min, temp_max in equipos:
        # Equipment header
        ws.merge_cells(f"A{current_row}:G{current_row}")
        ws.cell(row=current_row, column=1, value=f"📍 {equipo}  |  Rango: {temp_min}°C a {temp_max}°C").font = bold_font
        ws.cell(row=current_row, column=1).fill = blue_fill
        ws.cell(row=current_row, column=1).border = thin_border
        current_row += 1

        # Sub-headers
        headers = ["Día", "Temp. Mañana (°C)", "Estado", "Hora", "Temp. Tarde (°C)", "Estado", "Firma"]
        style_header_row(ws, current_row, 7, headers)
        current_row += 1

        # Data rows for 7 days
        for d_idx, dia in enumerate(dias):
            row = current_row + d_idx
            ws.cell(row=row, column=1, value=dia).font = data_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=1).alignment = left_align

            # Temp mañana (input)
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=2).number_format = '0.0'

            # Estado mañana (formula)
            formula = f'=IF(B{row}="","",IF(AND(B{row}>={temp_min},B{row}<={temp_max}),"OK","ALERTA"))'
            ws.cell(row=row, column=3, value=formula).font = data_font
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=3).alignment = center_align

            # Hora (input)
            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border

            # Temp tarde (input)
            ws.cell(row=row, column=5).fill = input_fill
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=5).number_format = '0.0'

            # Estado tarde (formula)
            formula_t = f'=IF(E{row}="","",IF(AND(E{row}>={temp_min},E{row}<={temp_max}),"OK","ALERTA"))'
            ws.cell(row=row, column=6, value=formula_t).font = data_font
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            # Firma
            ws.cell(row=row, column=7).fill = input_fill
            ws.cell(row=row, column=7).border = thin_border

        current_row += len(dias) + 1  # +1 spacer

    # Conditional formatting note
    add_footer(ws, current_row + 1, "Si aparece ALERTA: registrar incidencia en Hoja de Acciones Correctivas y avisar al responsable de calidad.")
    add_footer(ws, current_row + 2, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.print_area = f"A1:G{current_row + 2}"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "01-registro-temperaturas-diario.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 2: Registro de Temperaturas Recepción
# ══════════════════════════════════════════════════════════════
def create_temperaturas_recepcion():
    wb = Workbook()
    add_instructions_sheet(wb, "Control de Temperaturas en Recepción de Mercancías", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra la temperatura de cada producto al recibirlo",
        "▸ Compara con el límite máximo permitido",
        "▸ Si supera el límite: RECHAZAR o aplicar acción correctiva",
        "",
        "Límites de temperatura en recepción",
        "",
        "▸ Refrigerados (carne, pescado, lácteos): máx. 4°C",
        "▸ Congelados: máx. -18°C",
        "▸ Ultracongelados: máx. -20°C",
        "▸ Productos ambiente: sin control (verificar integridad envase)",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Recepción Temperaturas")
    ws.sheet_properties.tabColor = "FF9800"

    widths = {"A": 14, "B": 25, "C": 20, "D": 15, "E": 15, "F": 12, "G": 15, "H": 15, "I": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Control de Temperaturas — Recepción de Mercancías"
    ws["A1"].font = title_font

    ws["A2"].value = "Mes: _______________    Responsable: _________________________________"
    ws["A2"].font = subtitle_font

    headers = ["Fecha", "Producto", "Proveedor", "Temp. Recibida (°C)", "Límite Máx. (°C)", "Estado", "Lote / Caducidad", "Aceptado (S/N)", "Observaciones"]
    style_header_row(ws, 4, 9, headers)

    # Limits validation
    limit_validation = DataValidation(type="list", formula1='"4,-18,-20,N/A"', allow_blank=True)
    ws.add_data_validation(limit_validation)

    sn_validation = DataValidation(type="list", formula1='"S,N"', allow_blank=True)
    ws.add_data_validation(sn_validation)

    for i in range(30):
        row = 5 + i
        fill = light_row if i % 2 == 0 else alt_row
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = data_font
            if col in (1, 2, 3, 4, 7, 9):
                cell.fill = input_fill
            elif col == 5:
                cell.fill = input_fill
                limit_validation.add(cell)
            elif col == 8:
                cell.fill = input_fill
                sn_validation.add(cell)

        # Estado formula
        ws.cell(row=row, column=6, value=f'=IF(D{row}="","",IF(E{row}="N/A","OK",IF(D{row}<=E{row},"OK","RECHAZAR")))').border = thin_border
        ws.cell(row=row, column=6).alignment = center_align

    add_footer(ws, 36, "Si Estado = RECHAZAR: devolver producto al proveedor y registrar en Acciones Correctivas.")
    add_footer(ws, 37, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "02-registro-temperaturas-recepcion.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 3: Plan de Limpieza y Desinfección
# ══════════════════════════════════════════════════════════════
def create_plan_limpieza():
    wb = Workbook()
    add_instructions_sheet(wb, "Plan de Limpieza y Desinfección (L+D)", [
        "Cómo usar esta plantilla",
        "",
        "▸ Define qué se limpia, cuándo, cómo y quién lo hace",
        "▸ La pestaña 'Plan Maestro' es el documento oficial para inspección",
        "▸ Personaliza las zonas, productos y frecuencias de tu local",
        "▸ Imprime y coloca en zona visible de cocina",
        "",
        "Zonas estándar de un establecimiento",
        "",
        "▸ Cocina (suelos, superficies, equipos, cámaras, campana)",
        "▸ Sala/Comedor (mesas, sillas, barra, suelos)",
        "▸ Baños (inodoros, lavabos, espejos, suelos)",
        "▸ Almacén (estantes, suelos, contenedores)",
        "▸ Vestuarios (taquillas, duchas, bancos)",
        "▸ Exterior (terraza, entrada, contenedores basura)",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Plan Maestro L+D")
    ws.sheet_properties.tabColor = "4CAF50"

    widths = {"A": 25, "B": 30, "C": 15, "D": 25, "E": 20, "F": 25, "G": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Plan Maestro de Limpieza y Desinfección"
    ws["A1"].font = title_font

    ws["A2"].value = "Establecimiento: ________________________    Fecha revisión: ___/___/______"
    ws["A2"].font = subtitle_font

    headers = ["Zona / Elemento", "Qué se limpia", "Frecuencia", "Producto / Dilución", "Método", "Responsable", "Verificación"]
    style_header_row(ws, 4, 7, headers)

    freq_validation = DataValidation(type="list", formula1='"Cada servicio,Diaria,Semanal,Quincenal,Mensual,Trimestral"', allow_blank=True)
    ws.add_data_validation(freq_validation)

    # Pre-filled data
    zonas = [
        # (zona, que, frecuencia, producto, metodo, responsable, verificacion)
        ("COCINA", "", "", "", "", "", ""),
        ("Superficies de trabajo", "Encimeras, tablas de corte", "Cada servicio", "Desengrasante + desinfectante alimentario", "Pulverizar, frotar, aclarar, secar", "Cocinero de turno", "Visual + test superficie"),
        ("Suelos cocina", "Suelo completo", "Diaria", "Desengrasante diluido 1:20", "Barrer, fregar con fregona, secar", "Auxiliar cocina", "Visual"),
        ("Cámaras frigoríficas", "Interior, estantes, juntas puerta", "Semanal", "Desinfectante alimentario", "Vaciar, limpiar superficies, secar", "Jefe de cocina", "Visual + registro"),
        ("Campana extractora", "Filtros, superficie exterior", "Semanal filtros / Mensual campana", "Desengrasante industrial", "Desmontar filtros, sumergir, aclarar", "Empresa externa", "Registro limpieza"),
        ("Fogones y plancha", "Quemadores, plancha, horno", "Diaria", "Desengrasante cocina", "Rascar, pulverizar, frotar, aclarar", "Cocinero de turno", "Visual"),
        ("Fregaderos", "Pilas, grifos, sifones", "Diaria", "Desincrustante + lejía diluida", "Frotar con estropajo, aclarar", "Auxiliar cocina", "Visual"),
        ("Cubos de basura", "Interior y exterior cubos", "Diaria", "Lejía diluida 1:50", "Vaciar, lavar interior, secar", "Auxiliar cocina", "Visual + olor"),
        ("", "", "", "", "", "", ""),
        ("SALA / COMEDOR", "", "", "", "", "", ""),
        ("Mesas y sillas", "Superficie, patas", "Cada servicio", "Multiusos alimentario", "Pulverizar, frotar, secar", "Camarero", "Visual"),
        ("Barra", "Superficie, vitrinas, tiradores", "Cada servicio", "Desinfectante multiusos", "Pulverizar, frotar, secar", "Barman/camarero", "Visual"),
        ("Suelos sala", "Suelo completo", "Diaria", "Fregasuelos neutro", "Barrer, fregar, secar", "Personal limpieza", "Visual"),
        ("Cristalería/Vajilla", "Vasos, platos, cubiertos", "Cada uso", "Detergente lavavajillas + abrillantador", "Lavavajillas a 60°C mín.", "Personal cocina", "Visual + temp. lavavajillas"),
        ("", "", "", "", "", "", ""),
        ("BAÑOS", "", "", "", "", "", ""),
        ("Inodoros y urinarios", "Interior, exterior, cisterna", "2 veces/día", "Limpiador WC + lejía", "Aplicar, frotar con escobilla, aclarar", "Personal limpieza", "Visual + checklist"),
        ("Lavabos y espejos", "Pila, grifo, espejo, dispensadores", "2 veces/día", "Multiusos + limpiacristales", "Frotar, aclarar, secar", "Personal limpieza", "Visual"),
        ("Suelos baños", "Suelo completo", "2 veces/día", "Lejía diluida 1:50", "Fregar con fregona, secar", "Personal limpieza", "Visual + olor"),
        ("", "", "", "", "", "", ""),
        ("ALMACÉN", "", "", "", "", "", ""),
        ("Estantes y baldas", "Superficie estantes", "Semanal", "Multiusos", "Retirar producto, limpiar, reponer", "Almacenero", "Visual"),
        ("Suelos almacén", "Suelo completo", "Semanal", "Fregasuelos", "Barrer, fregar", "Personal limpieza", "Visual"),
        ("", "", "", "", "", "", ""),
        ("VESTUARIOS", "", "", "", "", "", ""),
        ("Taquillas y bancos", "Superficie exterior taquillas", "Semanal", "Multiusos", "Frotar, secar", "Personal limpieza", "Visual"),
        ("Duchas", "Plato ducha, paredes, grifo", "Diaria", "Antical + desinfectante", "Frotar, aclarar", "Personal limpieza", "Visual"),
    ]

    for i, (zona, que, freq, prod, metodo, resp, verif) in enumerate(zonas):
        row = 5 + i
        values = [zona, que, freq, prod, metodo, resp, verif]
        is_section = zona.isupper() and que == ""

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = left_align if col_idx <= 2 else center_align

            if is_section:
                cell.font = Font(name="Calibri", size=12, bold=True, color=GOLD)
                cell.fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
            elif val == "":
                cell.fill = input_fill
            else:
                cell.font = data_font
                cell.fill = light_row if i % 2 == 0 else alt_row

        if not is_section and freq != "":
            freq_validation.add(ws.cell(row=row, column=3))

    # Add 10 empty rows for user
    last = 5 + len(zonas)
    for i in range(10):
        row = last + i
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            if col == 3:
                freq_validation.add(cell)

    add_footer(ws, last + 12, "Este plan debe estar disponible para inspección sanitaria en todo momento.")
    add_footer(ws, last + 13, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "03-plan-limpieza-desinfeccion.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 4: Registro de Limpieza Diaria
# ══════════════════════════════════════════════════════════════
def create_registro_limpieza():
    wb = Workbook()
    add_instructions_sheet(wb, "Registro de Limpieza Diaria", [
        "Cómo usar esta plantilla",
        "",
        "▸ Imprime una hoja por semana",
        "▸ El responsable de limpieza marca con ✓ cada tarea completada",
        "▸ El encargado verifica y firma al final del día",
        "▸ Archiva las hojas firmadas — son obligatorias para inspección",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Limpieza Diaria")
    ws.sheet_properties.tabColor = "4CAF50"

    ws.column_dimensions["A"].width = 35
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Registro de Limpieza Diaria"
    ws["A1"].font = title_font

    ws["A2"].value = "Semana del: ___/___/______ al ___/___/______"
    ws["A2"].font = subtitle_font

    dias = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    headers = ["Tarea de Limpieza"] + dias
    style_header_row(ws, 4, 8, headers)

    tareas = [
        ("COCINA", True),
        ("Superficies de trabajo desinfectadas", False),
        ("Suelos barridos y fregados", False),
        ("Fogones y plancha limpiados", False),
        ("Fregaderos limpiados", False),
        ("Cubos de basura vaciados y limpiados", False),
        ("Cámaras: orden y limpieza exterior", False),
        ("Campana: filtros limpios (si toca)", False),
        ("", True),
        ("SALA / COMEDOR", True),
        ("Mesas y sillas limpiadas", False),
        ("Barra desinfectada", False),
        ("Suelos barridos y fregados", False),
        ("Cristalería repasada", False),
        ("", True),
        ("BAÑOS", True),
        ("Inodoros y lavabos limpiados", False),
        ("Espejos limpios", False),
        ("Suelos fregados", False),
        ("Jabón y papel repuestos", False),
        ("", True),
        ("VERIFICACIÓN", True),
        ("Verificado por (nombre):", False),
        ("Firma:", False),
    ]

    for i, (tarea, is_section) in enumerate(tareas):
        row = 5 + i

        cell_a = ws.cell(row=row, column=1, value=tarea)
        cell_a.border = thin_border
        cell_a.alignment = left_align

        if is_section and tarea != "":
            cell_a.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
            cell_a.fill = header_fill
            for c in range(2, 9):
                ws.cell(row=row, column=c).fill = header_fill
                ws.cell(row=row, column=c).border = thin_border
        elif tarea == "":
            pass
        else:
            cell_a.font = data_font
            for c in range(2, 9):
                cell = ws.cell(row=row, column=c)
                cell.fill = input_fill
                cell.border = thin_border
                cell.alignment = center_align

    add_footer(ws, 5 + len(tareas) + 2, "Marcar con ✓ cuando se complete. Archivar hojas firmadas durante mínimo 2 años.")
    add_footer(ws, 5 + len(tareas) + 3, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.page_setup.orientation = "portrait"
    ws.print_area = f"A1:H{5 + len(tareas)}"

    path = os.path.join(OUTPUT_DIR, "04-registro-limpieza-diaria.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 5: Checklist Recepción de Mercancías
# ══════════════════════════════════════════════════════════════
def create_checklist_recepcion():
    wb = Workbook()
    add_instructions_sheet(wb, "Checklist de Recepción de Mercancías", [
        "Cómo usar esta plantilla",
        "",
        "▸ Verifica CADA entrega de mercancía con este checklist",
        "▸ Controla: temperatura, caducidad, etiquetado, estado envase, cantidad",
        "▸ Si algo no cumple: RECHAZAR y registrar incidencia",
        "▸ Conserva los albaranes junto con este registro",
        "",
        "Puntos críticos de recepción",
        "",
        "▸ Temperatura: medir con termómetro de sonda (calibrado)",
        "▸ Caducidad: mínimo 2/3 de vida útil restante",
        "▸ Etiquetado: lote, origen, ingredientes, alérgenos",
        "▸ Envase: sin golpes, roturas ni hinchazón",
        "▸ Olor/aspecto: sin signos de deterioro",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Recepción Mercancías")
    ws.sheet_properties.tabColor = "FF9800"

    widths = {"A": 14, "B": 22, "C": 20, "D": 12, "E": 14, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12, "K": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:K1")
    ws["A1"].value = "Checklist de Recepción de Mercancías"
    ws["A1"].font = title_font

    ws["A2"].value = "Mes: _______________    Responsable: _________________________________"
    ws["A2"].font = subtitle_font

    headers = ["Fecha", "Producto", "Proveedor", "Temp. (°C)", "Caducidad", "Etiquetado OK", "Envase OK", "Aspecto OK", "Cantidad OK", "Aceptado", "Observaciones"]
    style_header_row(ws, 4, 11, headers)

    ok_validation = DataValidation(type="list", formula1='"✓,✗"', allow_blank=True)
    ws.add_data_validation(ok_validation)

    sn_validation = DataValidation(type="list", formula1='"SÍ,NO"', allow_blank=True)
    ws.add_data_validation(sn_validation)

    for i in range(25):
        row = 5 + i
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align
            cell.font = data_font

            if col in (6, 7, 8, 9):
                ok_validation.add(cell)
            if col == 10:
                sn_validation.add(cell)

    add_footer(ws, 31, "Conservar junto con albaranes de entrega. Archivo mínimo: 2 años.")
    add_footer(ws, 32, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "05-checklist-recepcion-mercancias.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 6: Registro de Trazabilidad
# ══════════════════════════════════════════════════════════════
def create_trazabilidad():
    wb = Workbook()
    add_instructions_sheet(wb, "Registro de Trazabilidad", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra TODOS los productos que entran y salen",
        "▸ Anota siempre: lote, proveedor, fecha entrada, fecha caducidad",
        "▸ En caso de alerta alimentaria, debes poder rastrear el origen en < 4 horas",
        "▸ Obligatorio según Reglamento (CE) 178/2002",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Trazabilidad")
    ws.sheet_properties.tabColor = "9C27B0"

    widths = {"A": 14, "B": 25, "C": 15, "D": 20, "E": 14, "F": 14, "G": 12, "H": 15, "I": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Registro de Trazabilidad — Entrada de Productos"
    ws["A1"].font = title_font

    ws["A2"].value = "Mes: _______________"
    ws["A2"].font = subtitle_font

    headers = ["Fecha Entrada", "Producto", "Nº Lote", "Proveedor", "Fecha Caducidad", "Fecha Consumo", "Cantidad", "Destino / Uso", "Observaciones"]
    style_header_row(ws, 4, 9, headers)

    for i in range(35):
        row = 5 + i
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align if col not in (2, 4, 8, 9) else left_align
            cell.font = data_font
            if col in (1, 5, 6):
                cell.number_format = 'DD/MM/YYYY'

    add_footer(ws, 41, "En caso de alerta alimentaria: localizar lote afectado, retirar producto, notificar a autoridad sanitaria.")
    add_footer(ws, 42, "Conservar registros mínimo 5 años (Reglamento CE 178/2002).")
    add_footer(ws, 43, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "06-registro-trazabilidad.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 7: Control de Plagas (DDD)
# ══════════════════════════════════════════════════════════════
def create_control_plagas():
    wb = Workbook()
    add_instructions_sheet(wb, "Registro de Control de Plagas (DDD)", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra TODAS las actuaciones de control de plagas",
        "▸ Incluye: desinsectación, desratización y desinfección",
        "▸ Guarda los certificados de la empresa de control de plagas",
        "▸ Actualiza el plano de ubicación de cebos/trampas",
        "",
        "Frecuencia recomendada",
        "",
        "▸ Desinsectación: trimestral (o según necesidad)",
        "▸ Desratización: mensual revisión de cebos",
        "▸ Desinfección general: semestral",
        "▸ Revisión visual: semanal (buscar indicios)",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Control Plagas DDD")
    ws.sheet_properties.tabColor = "795548"

    widths = {"A": 14, "B": 18, "C": 25, "D": 20, "E": 20, "F": 25, "G": 18, "H": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Registro de Control de Plagas (DDD)"
    ws["A1"].font = title_font

    ws["A2"].value = "Año: ______    Empresa contratada: ________________________________    Nº Contrato: ____________"
    ws["A2"].font = subtitle_font

    headers = ["Fecha", "Tipo Actuación", "Empresa / Técnico", "Productos Utilizados", "Zonas Tratadas", "Resultado / Hallazgos", "Próxima Revisión", "Nº Certificado"]
    style_header_row(ws, 4, 8, headers)

    tipo_validation = DataValidation(type="list", formula1='"Desinsectación,Desratización,Desinfección,Revisión cebos,Inspección visual,Otro"', allow_blank=True)
    ws.add_data_validation(tipo_validation)

    for i in range(20):
        row = 5 + i
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align if col in (1, 2, 7) else left_align
            cell.font = data_font
            if col == 2:
                tipo_validation.add(cell)
            if col in (1, 7):
                cell.number_format = 'DD/MM/YYYY'

    add_footer(ws, 26, "Adjuntar certificados de cada actuación. El plano de cebos debe estar actualizado y accesible.")
    add_footer(ws, 27, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "07-control-plagas-ddd.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 8: Matriz de Alérgenos
# ══════════════════════════════════════════════════════════════
def create_matriz_alergenos():
    wb = Workbook()
    add_instructions_sheet(wb, "Matriz de Alérgenos — Carta Completa", [
        "Cómo usar esta plantilla",
        "",
        "▸ Lista TODOS los platos de tu carta en la columna A",
        "▸ Para cada plato, marca los alérgenos presentes",
        "▸ Usa: S (Sí contiene), T (Trazas), N (No contiene)",
        "▸ Esta matriz es OBLIGATORIA (Reglamento UE 1169/2011)",
        "▸ Debe estar disponible para cualquier cliente que la solicite",
        "",
        "Los 14 alérgenos de declaración obligatoria",
        "",
        "▸ 1. Gluten | 2. Crustáceos | 3. Huevos | 4. Pescado",
        "▸ 5. Cacahuetes | 6. Soja | 7. Lácteos | 8. Frutos de cáscara",
        "▸ 9. Apio | 10. Mostaza | 11. Sésamo | 12. Sulfitos",
        "▸ 13. Altramuces | 14. Moluscos",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Matriz Alérgenos")
    ws.sheet_properties.tabColor = "E91E63"

    alergenos = [
        "Gluten", "Crustáceos", "Huevos", "Pescado", "Cacahuetes",
        "Soja", "Lácteos", "Frutos cáscara", "Apio", "Mostaza",
        "Sésamo", "Sulfitos", "Altramuces", "Moluscos"
    ]

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    for i in range(len(alergenos)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12

    ws.merge_cells(f"A1:{get_column_letter(4 + len(alergenos) - 1)}1")
    ws["A1"].value = "Matriz de Alérgenos — Carta Completa"
    ws["A1"].font = title_font

    ws["A2"].value = "Establecimiento: ________________________    Fecha actualización: ___/___/______"
    ws["A2"].font = subtitle_font

    ws["A3"].value = "Leyenda:  S = Contiene  |  T = Trazas  |  N = No contiene  |  (vacío) = No verificado"
    ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="555555")

    # Headers
    headers = ["Nº", "Plato / Producto", "Categoría"] + alergenos
    style_header_row(ws, 5, len(headers), headers)

    # Rotate allergen headers
    for i in range(len(alergenos)):
        cell = ws.cell(row=5, column=4 + i)
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=False)

    # Data validation
    stn_validation = DataValidation(type="list", formula1='"S,T,N"', allow_blank=True)
    ws.add_data_validation(stn_validation)

    cat_validation = DataValidation(type="list", formula1='"Entrantes,Primeros,Segundos,Postres,Bebidas,Tapas,Desayunos,Infantil,Otros"', allow_blank=True)
    ws.add_data_validation(cat_validation)

    # Example dishes
    ejemplo_platos = [
        ("Ensalada César", "Entrantes"),
        ("Croquetas de jamón", "Entrantes"),
        ("Sopa de pescado", "Primeros"),
        ("Paella mixta", "Primeros"),
        ("Solomillo con salsa Pedro Ximénez", "Segundos"),
        ("Merluza a la vasca", "Segundos"),
        ("Tarta de queso", "Postres"),
        ("Helado de vainilla", "Postres"),
    ]

    for i in range(40):
        row = 6 + i
        fill = light_row if i % 2 == 0 else alt_row

        # Nº
        ws.cell(row=row, column=1, value=i + 1 if i < len(ejemplo_platos) else "").font = data_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align

        # Plato
        plato = ejemplo_platos[i][0] if i < len(ejemplo_platos) else ""
        ws.cell(row=row, column=2, value=plato).font = data_font
        ws.cell(row=row, column=2).fill = input_fill
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = left_align

        # Categoría
        cat = ejemplo_platos[i][1] if i < len(ejemplo_platos) else ""
        ws.cell(row=row, column=3, value=cat).font = data_font
        ws.cell(row=row, column=3).fill = input_fill
        ws.cell(row=row, column=3).border = thin_border
        cat_validation.add(ws.cell(row=row, column=3))

        # Allergen cells
        for j in range(len(alergenos)):
            cell = ws.cell(row=row, column=4 + j)
            cell.fill = input_fill
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = data_font
            stn_validation.add(cell)

    # Count row
    count_row = 47
    ws.cell(row=count_row, column=2, value="TOTAL platos con alérgeno:").font = bold_font
    ws.cell(row=count_row, column=2).border = thin_border
    for j in range(len(alergenos)):
        col = 4 + j
        formula = f'=COUNTIF({get_column_letter(col)}6:{get_column_letter(col)}45,"S")+COUNTIF({get_column_letter(col)}6:{get_column_letter(col)}45,"T")'
        ws.cell(row=count_row, column=col, value=formula).font = bold_font
        ws.cell(row=count_row, column=col).border = thin_border
        ws.cell(row=count_row, column=col).alignment = center_align

    add_footer(ws, 49, "Esta carta de alérgenos debe estar disponible para cualquier cliente que la solicite (Reglamento UE 1169/2011).")
    add_footer(ws, 50, "Actualizar cada vez que se modifique la carta o los ingredientes de un plato.")
    add_footer(ws, 51, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "D6"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "08-matriz-alergenos.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 9: Control de Aceite de Fritura
# ══════════════════════════════════════════════════════════════
def create_control_aceite():
    wb = Workbook()
    add_instructions_sheet(wb, "Control de Aceite de Fritura", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra cada cambio de aceite y resultado de test",
        "▸ El aceite debe cambiarse cuando supera el 25% de compuestos polares",
        "▸ Temperatura máxima de fritura: 180°C",
        "▸ Usa tiras reactivas o medidor digital de compuestos polares",
        "▸ Obligatorio según RD 2207/1995",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Control Aceite")
    ws.sheet_properties.tabColor = "FFC107"

    widths = {"A": 14, "B": 20, "C": 18, "D": 18, "E": 15, "F": 12, "G": 18, "H": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Control de Aceite de Fritura"
    ws["A1"].font = title_font

    ws["A2"].value = "Freidora: ________________________    Tipo aceite: ________________________"
    ws["A2"].font = subtitle_font

    headers = ["Fecha", "Freidora / Equipo", "Tipo de Test", "Resultado (%CP)", "Temp. Máx (°C)", "Estado", "Acción Realizada", "Firma"]
    style_header_row(ws, 4, 8, headers)

    test_validation = DataValidation(type="list", formula1='"Tiras reactivas,Medidor digital,Visual,Otro"', allow_blank=True)
    ws.add_data_validation(test_validation)

    for i in range(20):
        row = 5 + i
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align
            cell.font = data_font
            if col == 3:
                test_validation.add(cell)

        # Estado formula: if >25% = CAMBIAR, if >20% = VIGILAR, else OK
        ws.cell(row=row, column=6, value=f'=IF(D{row}="","",IF(D{row}>25,"CAMBIAR",IF(D{row}>20,"VIGILAR","OK")))').border = thin_border
        ws.cell(row=row, column=6).alignment = center_align

    add_footer(ws, 26, "Límite legal: máx. 25% compuestos polares (RD 2207/1995). Cambiar aceite antes de superar este límite.")
    add_footer(ws, 27, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "09-control-aceite-fritura.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 10: Control de Agua
# ══════════════════════════════════════════════════════════════
def create_control_agua():
    wb = Workbook()
    add_instructions_sheet(wb, "Registro de Control de Agua Potable", [
        "Cómo usar esta plantilla",
        "",
        "▸ Si tu agua viene de la red municipal: registro simplificado",
        "▸ Si tienes pozo o depósito propio: análisis obligatorios",
        "▸ Registra niveles de cloro residual (0.2-1.0 mg/L)",
        "▸ Guarda los boletines de análisis de la empresa de agua",
        "▸ Obligatorio según RD 140/2003",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Control Agua")
    ws.sheet_properties.tabColor = "03A9F4"

    widths = {"A": 14, "B": 18, "C": 18, "D": 18, "E": 15, "F": 20, "G": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Registro de Control de Agua Potable"
    ws["A1"].font = title_font

    ws["A2"].value = "Fuente de suministro: ☐ Red municipal   ☐ Pozo propio   ☐ Depósito"
    ws["A2"].font = subtitle_font

    headers = ["Fecha", "Punto de Muestreo", "Cloro Residual (mg/L)", "Aspecto / Olor", "Estado", "Análisis Externo (S/N)", "Observaciones"]
    style_header_row(ws, 4, 7, headers)

    aspecto_validation = DataValidation(type="list", formula1='"Normal,Turbio,Olor extraño,Color anormal"', allow_blank=True)
    ws.add_data_validation(aspecto_validation)

    for i in range(15):
        row = 5 + i
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align
            cell.font = data_font
            if col == 4:
                aspecto_validation.add(cell)

        # Estado formula
        ws.cell(row=row, column=5, value=f'=IF(C{row}="","",IF(AND(C{row}>=0.2,C{row}<=1.0,D{row}="Normal"),"OK","REVISAR"))').border = thin_border

    add_footer(ws, 21, "Cloro residual aceptable: 0.2 - 1.0 mg/L (RD 140/2003).")
    add_footer(ws, 22, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "10-control-agua-potable.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 11: Registro de Acciones Correctivas
# ══════════════════════════════════════════════════════════════
def create_acciones_correctivas():
    wb = Workbook()
    add_instructions_sheet(wb, "Registro de Acciones Correctivas", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra CUALQUIER incidencia relacionada con seguridad alimentaria",
        "▸ Incluye: desviaciones de temperatura, producto rechazado, reclamaciones, etc.",
        "▸ Documenta la causa, la acción tomada y la verificación",
        "▸ Este registro demuestra al inspector que gestionas las incidencias",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Acciones Correctivas")
    ws.sheet_properties.tabColor = "F44336"

    widths = {"A": 8, "B": 14, "C": 20, "D": 30, "E": 20, "F": 30, "G": 15, "H": 14, "I": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Registro de Acciones Correctivas"
    ws["A1"].font = title_font

    ws["A2"].value = "Año: ______"
    ws["A2"].font = subtitle_font

    headers = ["Nº", "Fecha", "Tipo Incidencia", "Descripción del Problema", "Causa Probable", "Acción Correctiva Adoptada", "Responsable", "Fecha Cierre", "Verificación"]
    style_header_row(ws, 4, 9, headers)

    tipo_validation = DataValidation(type="list", formula1='"Temperatura,Producto rechazado,Limpieza,Plagas,Reclamación cliente,Alérgeno,Contaminación,Caducidad,Otro"', allow_blank=True)
    ws.add_data_validation(tipo_validation)

    verif_validation = DataValidation(type="list", formula1='"Pendiente,Verificado OK,Requiere seguimiento"', allow_blank=True)
    ws.add_data_validation(verif_validation)

    for i in range(25):
        row = 5 + i
        ws.cell(row=row, column=1, value=i + 1).font = data_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align

        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = left_align if col in (4, 5, 6) else center_align
            cell.font = data_font
            if col == 3:
                tipo_validation.add(cell)
            if col == 9:
                verif_validation.add(cell)
            if col in (2, 8):
                cell.number_format = 'DD/MM/YYYY'

    add_footer(ws, 31, "Toda incidencia debe tener una acción correctiva documentada y verificada.")
    add_footer(ws, 32, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "11-registro-acciones-correctivas.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 12: Análisis de Peligros HACCP
# ══════════════════════════════════════════════════════════════
def create_analisis_haccp():
    wb = Workbook()
    add_instructions_sheet(wb, "Análisis de Peligros HACCP", [
        "Cómo usar esta plantilla",
        "",
        "▸ Identifica los peligros en cada fase del proceso alimentario",
        "▸ Clasifica: Biológico (B), Químico (Q), Físico (F)",
        "▸ Determina los Puntos Críticos de Control (PCC)",
        "▸ Define límites críticos, sistema de vigilancia y acciones correctivas",
        "▸ Este documento es el corazón del sistema APPCC",
        "",
        "Los 7 principios HACCP",
        "",
        "▸ 1. Análisis de peligros",
        "▸ 2. Determinación de los PCC",
        "▸ 3. Establecimiento de límites críticos",
        "▸ 4. Sistema de vigilancia de los PCC",
        "▸ 5. Medidas correctivas",
        "▸ 6. Procedimientos de verificación",
        "▸ 7. Sistema de documentación y registro",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Análisis Peligros")
    ws.sheet_properties.tabColor = "FF5722"

    widths = {"A": 22, "B": 12, "C": 25, "D": 12, "E": 12, "F": 20, "G": 20, "H": 20, "I": 20, "J": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Análisis de Peligros — Plan HACCP"
    ws["A1"].font = title_font

    ws["A2"].value = "Establecimiento: ________________________    Fecha: ___/___/______    Revisión nº: _____"
    ws["A2"].font = subtitle_font

    headers = ["Fase del Proceso", "Tipo Peligro", "Peligro Identificado", "Probabilidad", "Gravedad", "¿Es PCC?", "Límite Crítico", "Vigilancia", "Acción Correctiva", "Registro"]
    style_header_row(ws, 4, 10, headers)

    tipo_validation = DataValidation(type="list", formula1='"B (Biológico),Q (Químico),F (Físico)"', allow_blank=True)
    ws.add_data_validation(tipo_validation)

    prob_validation = DataValidation(type="list", formula1='"Alta,Media,Baja"', allow_blank=True)
    ws.add_data_validation(prob_validation)

    sn_validation = DataValidation(type="list", formula1='"SÍ (PCC),NO"', allow_blank=True)
    ws.add_data_validation(sn_validation)

    # Pre-filled HACCP analysis
    fases = [
        ("RECEPCIÓN", "B (Biológico)", "Temperatura elevada en producto refrigerado", "Alta", "Alta", "SÍ (PCC)", "≤ 4°C refrigerados\n≤ -18°C congelados", "Medir temp. cada recepción", "Rechazar producto", "Reg. Temperaturas Recepción"),
        ("RECEPCIÓN", "Q (Químico)", "Producto sin etiquetado de alérgenos", "Media", "Alta", "SÍ (PCC)", "Etiquetado completo", "Verificar etiqueta", "Rechazar producto", "Checklist Recepción"),
        ("RECEPCIÓN", "F (Físico)", "Envase dañado / presencia cuerpos extraños", "Baja", "Media", "NO", "Envase íntegro", "Inspección visual", "Rechazar producto", "Checklist Recepción"),
        ("ALMACENAMIENTO", "B (Biológico)", "Rotura cadena de frío", "Media", "Alta", "SÍ (PCC)", "0-4°C cámara\n-18°C congelador", "Registro temp. 2x/día", "Evaluar producto / desechar", "Reg. Temperaturas Diario"),
        ("ALMACENAMIENTO", "B (Biológico)", "Contaminación cruzada en cámara", "Media", "Alta", "NO", "Separación crudo/cocinado", "Inspección visual", "Reorganizar cámara", "Checklist semanal"),
        ("ALMACENAMIENTO", "Q (Químico)", "Producto caducado en stock", "Media", "Media", "NO", "FIFO / fecha caducidad", "Revisión semanal", "Retirar y registrar", "Reg. Trazabilidad"),
        ("PREPARACIÓN", "B (Biológico)", "Contaminación cruzada por utensilios", "Media", "Alta", "NO", "Tablas/cuchillos por color", "Supervisión continua", "Lavar y desinfectar", "Plan L+D"),
        ("PREPARACIÓN", "F (Físico)", "Restos metálicos de estropajo", "Baja", "Alta", "NO", "Uso estropajos alimentarios", "Revisión estado utensilios", "Sustituir estropajo", "Checklist limpieza"),
        ("COCCIÓN", "B (Biológico)", "Temperatura insuficiente en centro", "Media", "Alta", "SÍ (PCC)", "≥ 75°C centro producto\n(aves ≥ 82°C)", "Medir temp. centro", "Continuar cocción", "Registro cocción"),
        ("MANTENIMIENTO CALIENTE", "B (Biológico)", "Temperatura < 65°C", "Media", "Alta", "SÍ (PCC)", "≥ 65°C", "Medir temp. periódica", "Recalentar > 75°C o desechar", "Reg. Temperaturas"),
        ("ENFRIAMIENTO", "B (Biológico)", "Enfriamiento lento (zona peligro)", "Media", "Alta", "SÍ (PCC)", "De 60°C a 10°C\nen < 2 horas", "Medir tiempo/temp.", "Desechar si > 2h", "Registro enfriamiento"),
        ("SERVICIO", "B (Biológico)", "Exposición prolongada a temp. ambiente", "Media", "Media", "NO", "Máx. 2h a temp. ambiente", "Control horario", "Retirar alimento", "Registro servicio"),
        ("SERVICIO", "Q (Químico)", "Alérgeno no declarado al cliente", "Media", "Alta", "SÍ (PCC)", "Carta alérgenos actualizada", "Consultar carta alérgenos", "Protocolo alergia", "Matriz Alérgenos"),
    ]

    for i, (fase, tipo, peligro, prob, grav, pcc, limite, vigil, accion, reg) in enumerate(fases):
        row = 5 + i
        values = [fase, tipo, peligro, prob, grav, pcc, limite, vigil, accion, reg]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = data_font
            cell.alignment = left_align if col_idx in (3, 7, 8, 9, 10) else center_align

            if "SÍ" in str(val):
                cell.font = Font(name="Calibri", size=11, bold=True, color=RED_ALERT)
            if fase.isupper() and col_idx == 1 and (i == 0 or fases[i-1][0] != fase):
                cell.font = bold_font

    # Empty rows for user
    for i in range(10):
        row = 5 + len(fases) + i
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align
            if col == 2:
                tipo_validation.add(cell)
            if col in (4, 5):
                prob_validation.add(cell)
            if col == 6:
                sn_validation.add(cell)

    last = 5 + len(fases) + 10
    add_footer(ws, last + 1, "PCC = Punto Crítico de Control. Requiere vigilancia continua, límite crítico y acción correctiva definida.")
    add_footer(ws, last + 2, "Revisar este análisis al menos 1 vez al año o cuando cambie el menú/proceso.")
    add_footer(ws, last + 3, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "12-analisis-peligros-haccp.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 13: Checklist Higiene Personal (imprimible)
# ══════════════════════════════════════════════════════════════
def create_higiene_personal():
    wb = Workbook()
    add_instructions_sheet(wb, "Checklist de Higiene Personal", [
        "Cómo usar esta plantilla",
        "",
        "▸ Imprime y coloca en vestuario y entrada a cocina",
        "▸ Todo el personal debe cumplir ANTES de entrar a zona de manipulación",
        "▸ El encargado verifica semanalmente",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Higiene Personal")
    ws.sheet_properties.tabColor = "00BCD4"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12

    ws.merge_cells("B2:C2")
    ws["B2"].value = "NORMAS DE HIGIENE PERSONAL — Zona de Manipulación"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=RED_ALERT)

    normas = [
        ("INDUMENTARIA", True),
        ("Uniforme limpio y exclusivo para trabajo", False),
        ("Gorro o cofia que cubra todo el cabello", False),
        ("Calzado cerrado, antideslizante y limpio", False),
        ("Sin joyas, relojes ni piercings visibles", False),
        ("Delantal limpio (cambiar si se ensucia)", False),
        ("", False),
        ("MANOS", True),
        ("Uñas cortas, limpias, sin esmalte", False),
        ("Lavado de manos al entrar a cocina", False),
        ("Lavado de manos después de ir al baño", False),
        ("Lavado de manos al cambiar de tarea", False),
        ("Lavado de manos después de tocar basura", False),
        ("Uso de guantes cuando sea necesario (heridas)", False),
        ("Cubrir heridas con apósito azul impermeable", False),
        ("", False),
        ("CONDUCTA", True),
        ("No comer, beber ni fumar en zona de trabajo", False),
        ("No toser ni estornudar sobre alimentos", False),
        ("No tocar cara, pelo ni nariz mientras se cocina", False),
        ("Informar al encargado si se está enfermo (vómitos, diarrea, fiebre)", False),
        ("No trabajar con síntomas gastrointestinales", False),
        ("", False),
        ("FORMACIÓN", True),
        ("Carné/certificado de manipulador de alimentos vigente", False),
        ("Formación APPCC básica recibida", False),
        ("Conocimiento de alérgenos y protocolo", False),
    ]

    for i, (norma, is_section) in enumerate(normas):
        row = 4 + i
        cell_b = ws.cell(row=row, column=2, value=norma)
        cell_b.border = thin_border

        if is_section:
            cell_b.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
            cell_b.fill = header_fill
            ws.cell(row=row, column=3).fill = header_fill
            ws.cell(row=row, column=3).border = thin_border
        elif norma == "":
            pass
        else:
            cell_b.font = data_font
            cell_b.alignment = left_align
            cell_c = ws.cell(row=row, column=3, value="☐")
            cell_c.font = Font(name="Calibri", size=14)
            cell_c.alignment = center_align
            cell_c.border = thin_border

    last = 4 + len(normas)
    ws.cell(row=last + 2, column=2, value="Responsable verificación: _________________________    Fecha: ___/___/______").font = subtitle_font

    add_footer(ws, last + 4, "Imprimir y colocar en vestuario y puerta de acceso a cocina.")
    add_footer(ws, last + 5, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.page_setup.orientation = "portrait"

    path = os.path.join(OUTPUT_DIR, "13-checklist-higiene-personal.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 14: Fichas de los 14 Alérgenos (imprimible)
# ══════════════════════════════════════════════════════════════
def create_fichas_alergenos():
    wb = Workbook()
    add_instructions_sheet(wb, "Fichas de los 14 Alérgenos Obligatorios", [
        "Cómo usar esta plantilla",
        "",
        "▸ Imprime estas fichas y colócalas en cocina y sala",
        "▸ Todo el personal debe conocer los 14 alérgenos",
        "▸ Obligatorio según Reglamento UE 1169/2011 (RD 126/2015 en España)",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    alergenos_data = [
        ("1. Gluten", "Cereales con gluten: trigo, centeno, cebada, avena, espelta, kamut y sus derivados", "Pan, pasta, bollería, rebozados, cerveza, salsas con harina, cuscús, seitan"),
        ("2. Crustáceos", "Cangrejos, langostinos, gambas, bogavante, nécora, cigalas y derivados", "Paella, salpicón de marisco, croquetas de marisco, sopas de pescado, surimi"),
        ("3. Huevos", "Huevos y derivados de cualquier ave", "Tortillas, rebozados, mayonesa, pasta fresca, merengue, bizcochos, helados, flanes"),
        ("4. Pescado", "Todas las especies de pescado y derivados", "Paella, caldos, sopas, surimi, salsa Worcestershire, gelatina de pescado"),
        ("5. Cacahuetes", "Cacahuetes y derivados", "Salsas asiáticas, aceite de cacahuete, mantequilla de cacahuete, snacks, postres"),
        ("6. Soja", "Semillas de soja y derivados", "Salsa de soja, tofu, lecitina (E322), aceite de soja, miso, tempeh, edamame"),
        ("7. Lácteos", "Leche de vaca, cabra, oveja y derivados (incluida lactosa)", "Quesos, nata, mantequilla, yogur, bechamel, helados, chocolate con leche"),
        ("8. Frutos de cáscara", "Almendras, avellanas, nueces, anacardos, pecanas, pistachos, macadamias, nuez de Brasil", "Postres, pralinés, pesto, turrón, mazapán, aceites, mantequillas de frutos secos"),
        ("9. Apio", "Apio y derivados (incluido sal de apio)", "Sopas, caldos, ensaladas, snacks, salsas, especias mixtas"),
        ("10. Mostaza", "Semillas de mostaza y derivados", "Salsas, vinagretas, encurtidos, marinados, especias, curry en polvo"),
        ("11. Sésamo", "Semillas de sésamo y derivados", "Pan de hamburguesa, hummus, tahini, aceite de sésamo, sushi, ensaladas"),
        ("12. Sulfitos", "Dióxido de azufre y sulfitos (> 10 mg/kg o 10 mg/L)", "Vino, cerveza, vinagre, frutas desecadas, crustáceos, patatas procesadas"),
        ("13. Altramuces", "Altramuces y derivados", "Harina de altramuz en pan/bollería, snacks, sustitutos de café"),
        ("14. Moluscos", "Mejillones, almejas, ostras, calamares, pulpo, caracoles y derivados", "Paella, fideuá, tapas, ensaladas de marisco, tinta de calamar"),
    ]

    ws = wb.create_sheet(title="14 Alérgenos")
    ws.sheet_properties.tabColor = "E91E63"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50

    ws.merge_cells("B1:D1")
    ws["B1"].value = "LOS 14 ALÉRGENOS DE DECLARACIÓN OBLIGATORIA"
    ws["B1"].font = Font(name="Calibri", size=16, bold=True, color=RED_ALERT)

    ws["B2"].value = "Reglamento UE 1169/2011 — Real Decreto 126/2015"
    ws["B2"].font = subtitle_font

    headers = ["Nº", "Alérgeno", "Descripción", "Dónde se encuentra frecuentemente"]
    style_header_row(ws, 4, 4, headers)

    for i, (nombre, desc, donde) in enumerate(alergenos_data):
        row = 5 + i
        fill = light_row if i % 2 == 0 else alt_row

        ws.cell(row=row, column=1, value=i + 1).font = bold_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align

        ws.cell(row=row, column=2, value=nombre.split(". ")[1]).font = Font(name="Calibri", size=11, bold=True, color=RED_ALERT)
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = left_align

        ws.cell(row=row, column=3, value=desc).font = data_font
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = left_align

        ws.cell(row=row, column=4, value=donde).font = data_font
        ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = left_align

    last = 5 + len(alergenos_data)
    ws.merge_cells(f"B{last+1}:D{last+1}")
    ws.cell(row=last+1, column=2, value="PROTOCOLO: Si un cliente informa de alergia → consultar carta de alérgenos → informar al chef → preparar con utensilios limpios → servir con aviso al camarero").font = Font(name="Calibri", size=10, bold=True, color=RED_ALERT)

    add_footer(ws, last + 3, "Imprimir y colocar en zona visible de cocina y barra. Todo el personal debe conocer estos alérgenos.")
    add_footer(ws, last + 4, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "14-fichas-14-alergenos.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# BONUS 1: Registro de Formación en Seguridad Alimentaria
# ══════════════════════════════════════════════════════════════
def create_formacion():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS: Registro de Formación en Seguridad Alimentaria", [
        "Cómo usar esta plantilla",
        "",
        "▸ Registra TODA la formación del personal en seguridad alimentaria",
        "▸ Incluye: manipulador de alimentos, APPCC, alérgenos, primeros auxilios",
        "▸ El inspector puede pedir este registro en cualquier inspección",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Formación Personal")
    ws.sheet_properties.tabColor = GOLD

    widths = {"A": 22, "B": 15, "C": 30, "D": 14, "E": 14, "F": 14, "G": 20, "H": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    ws["A1"].value = "BONUS — Registro de Formación en Seguridad Alimentaria"
    ws["A1"].font = title_font

    headers = ["Nombre Empleado", "Puesto", "Formación Recibida", "Fecha Inicio", "Fecha Fin", "Duración (h)", "Entidad Formadora", "Nº Certificado"]
    style_header_row(ws, 3, 8, headers)

    tipo_formacion = DataValidation(type="list", formula1='"Manipulador alimentos,APPCC básico,APPCC avanzado,Alérgenos,Primeros auxilios,Prevención riesgos,Atención al cliente,Otro"', allow_blank=True)
    ws.add_data_validation(tipo_formacion)

    for i in range(20):
        row = 4 + i
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = center_align if col in (4, 5, 6) else left_align
            cell.font = data_font
            if col == 3:
                tipo_formacion.add(cell)
            if col in (4, 5):
                cell.number_format = 'DD/MM/YYYY'

    add_footer(ws, 25, "Todo el personal que manipula alimentos debe tener formación acreditada vigente.")
    add_footer(ws, 26, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "BONUS-01-registro-formacion.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# BONUS 2: Protocolo Alerta Alimentaria (imprimible)
# ══════════════════════════════════════════════════════════════
def create_protocolo_alerta():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS: Protocolo de Actuación ante Alerta Alimentaria", [
        "Cómo usar esta plantilla",
        "",
        "▸ Imprime en tamaño A3 y coloca en zona visible de cocina",
        "▸ Todo el personal debe conocer estos pasos",
        "▸ Actuar rápido puede evitar sanciones y proteger al cliente",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Protocolo Alerta")
    ws.sheet_properties.tabColor = RED_ALERT

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 35

    ws.merge_cells("B2:C2")
    ws["B2"].value = "PROTOCOLO DE ACTUACIÓN ANTE ALERTA ALIMENTARIA"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=RED_ALERT)

    ws["B3"].value = "Seguir estos pasos en orden. No improvisar."
    ws["B3"].font = Font(name="Calibri", size=12, bold=True, color="555555")

    pasos = [
        ("PASO 1: IDENTIFICAR", "Identificar el producto afectado (nombre, lote, proveedor, fecha recepción)", "Consultar registro de trazabilidad"),
        ("PASO 2: AISLAR", "Retirar INMEDIATAMENTE el producto de la cadena alimentaria", "Marcar como 'NO USAR — ALERTA' y separar en zona aislada"),
        ("PASO 3: NOTIFICAR", "Avisar al responsable de calidad / gerente del establecimiento", "Si hay riesgo para la salud: llamar al 112 / Sanidad"),
        ("PASO 4: DOCUMENTAR", "Registrar: qué producto, cuánto hay en stock, a quién se ha servido", "Completar hoja de Acciones Correctivas"),
        ("PASO 5: COMUNICAR", "Si se ha servido al público: evaluar si hay que informar a los clientes", "Contactar con el proveedor para coordinar retirada"),
        ("PASO 6: VERIFICAR", "Comprobar que no queda producto afectado en ningún punto del local", "Revisar cámaras, almacén, sala, preparaciones en curso"),
        ("PASO 7: REGISTRAR", "Documentar todo el proceso: fechas, acciones, personas implicadas", "Archivar junto con el registro de acciones correctivas"),
    ]

    current_row = 5
    for paso_title, accion, detalle in pasos:
        # Title row
        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws.cell(row=current_row, column=2, value=paso_title).font = Font(name="Calibri", size=13, bold=True, color=WHITE)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=RED_ALERT, end_color=RED_ALERT, fill_type="solid")
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=RED_ALERT, end_color=RED_ALERT, fill_type="solid")
        ws.cell(row=current_row, column=3).border = thin_border
        current_row += 1

        # Action row
        ws.cell(row=current_row, column=2, value=accion).font = data_font
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = left_align

        ws.cell(row=current_row, column=3, value=detalle).font = Font(name="Calibri", size=11, color="555555")
        ws.cell(row=current_row, column=3).border = thin_border
        ws.cell(row=current_row, column=3).alignment = left_align
        current_row += 2  # spacer

    # Emergency contacts
    current_row += 1
    ws.merge_cells(f"B{current_row}:C{current_row}")
    ws.cell(row=current_row, column=2, value="TELÉFONOS DE EMERGENCIA").font = Font(name="Calibri", size=14, bold=True, color=GOLD)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=3).fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    ws.cell(row=current_row, column=3).border = thin_border

    contactos = [
        ("Emergencias generales", "112"),
        ("Centro de Salud más cercano", "(completar)"),
        ("Responsable de calidad", "(nombre + teléfono)"),
        ("Empresa control de plagas", "(nombre + teléfono)"),
        ("Proveedor principal", "(nombre + teléfono)"),
    ]

    for nombre, tel in contactos:
        current_row += 1
        ws.cell(row=current_row, column=2, value=nombre).font = bold_font
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3, value=tel).font = data_font
        ws.cell(row=current_row, column=3).fill = input_fill
        ws.cell(row=current_row, column=3).border = thin_border

    add_footer(ws, current_row + 3, "Imprimir en A3 y colocar en zona visible. Todo el personal debe conocer este protocolo.")
    add_footer(ws, current_row + 4, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.page_setup.orientation = "portrait"

    path = os.path.join(OUTPUT_DIR, "BONUS-02-protocolo-alerta-alimentaria.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# TEMPLATE 15: Guía Inspección Sanidad (content sheet)
# ══════════════════════════════════════════════════════════════
def create_guia_inspeccion():
    wb = Workbook()
    add_instructions_sheet(wb, "Guía: Cómo Pasar una Inspección de Sanidad", [
        "Sobre esta guía",
        "",
        "▸ Los 25 puntos que revisa el inspector de Sanidad",
        "▸ Qué errores causan sanción inmediata",
        "▸ Cómo prepararte antes de la inspección",
        "▸ Qué documentos tener listos",
        "▸ Basada en la experiencia de 15 años de consultoría gastronómica",
        "",
        "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="25 Puntos Inspección")
    ws.sheet_properties.tabColor = "F44336"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Los 25 Puntos que Revisa el Inspector de Sanidad"
    ws["A1"].font = title_font

    ws["A2"].value = "Usa esta guía para autoevaluar tu establecimiento ANTES de una inspección"
    ws["A2"].font = subtitle_font

    headers = ["Nº", "Punto de Inspección", "Qué Revisa el Inspector", "Gravedad", "Tu Estado"]
    style_header_row(ws, 4, 5, headers)

    estado_validation = DataValidation(type="list", formula1='"✓ Cumple,⚠ Mejorar,✗ No cumple"', allow_blank=True)
    ws.add_data_validation(estado_validation)

    puntos = [
        ("DOCUMENTACIÓN", "", "", ""),
        ("Plan APPCC completo y actualizado", "Análisis de peligros, PCC, vigilancia, acciones correctivas", "GRAVE", ""),
        ("Registros de control al día", "Temperaturas, limpieza, trazabilidad, acciones correctivas", "GRAVE", ""),
        ("Certificados de formación del personal", "Manipulador de alimentos de todos los empleados", "GRAVE", ""),
        ("Carta de alérgenos accesible", "Disponible para cualquier cliente que la solicite", "GRAVE", ""),
        ("Contrato empresa DDD (plagas)", "Vigente, con certificados de actuaciones", "MODERADA", ""),
        ("", "", "", ""),
        ("TEMPERATURAS", "", "", ""),
        ("Cámaras frigoríficas (0-4°C)", "Medición in situ con termómetro del inspector", "GRAVE", ""),
        ("Congeladores (≤ -18°C)", "Medición in situ + productos congelados correctamente", "GRAVE", ""),
        ("Exposición caliente (≥ 65°C)", "Temperatura de mantenimiento de platos calientes", "GRAVE", ""),
        ("Termómetro calibrado disponible", "Sonda limpia, calibrada, accesible", "MODERADA", ""),
        ("", "", "", ""),
        ("HIGIENE Y LIMPIEZA", "", "", ""),
        ("Limpieza general del local", "Suelos, paredes, techos, equipos, sin grasa acumulada", "MODERADA", ""),
        ("Estado de superficies de trabajo", "Sin grietas, desperfectos, materiales aptos uso alimentario", "MODERADA", ""),
        ("Lavamanos en cocina", "Con agua caliente, jabón, papel secamanos, papelera con pedal", "GRAVE", ""),
        ("Separación zonas sucias/limpias", "Flujo de trabajo sin cruces entre crudo y cocinado", "GRAVE", ""),
        ("Vestuarios del personal", "Separación ropa de calle/trabajo, taquillas, limpieza", "LEVE", ""),
        ("", "", "", ""),
        ("ALMACENAMIENTO", "", "", ""),
        ("FIFO respetado", "Producto más antiguo delante, sin caducados", "GRAVE", ""),
        ("Separación crudo/cocinado", "Nunca en el mismo estante, crudo siempre abajo", "GRAVE", ""),
        ("Etiquetado de producto", "Todo identificado: nombre, fecha elaboración, caducidad", "MODERADA", ""),
        ("Producto no en suelo", "Todo en estanterías, nada directamente en el suelo", "MODERADA", ""),
        ("Productos de limpieza separados", "Nunca junto a alimentos, en armario cerrado", "GRAVE", ""),
        ("", "", "", ""),
        ("MANIPULACIÓN Y SERVICIO", "", "", ""),
        ("Uniforme del personal", "Limpio, gorro, sin joyas, calzado cerrado", "MODERADA", ""),
        ("Aceite de fritura", "Test de compuestos polares, registro de cambios", "MODERADA", ""),
        ("Descongelación correcta", "En cámara o microondas, nunca a temperatura ambiente", "GRAVE", ""),
        ("Control de alérgenos en servicio", "Personal formado, protocolo ante declaración de alergia", "GRAVE", ""),
    ]

    current_row = 5
    num = 0
    for punto, revision, gravedad, estado in puntos:
        if punto.isupper() and revision == "":
            # Section header
            ws.merge_cells(f"A{current_row}:E{current_row}")
            ws.cell(row=current_row, column=1, value=punto).font = Font(name="Calibri", size=12, bold=True, color=WHITE)
            ws.cell(row=current_row, column=1).fill = header_fill
            ws.cell(row=current_row, column=1).border = thin_border
            for c in range(2, 6):
                ws.cell(row=current_row, column=c).fill = header_fill
                ws.cell(row=current_row, column=c).border = thin_border
        elif punto == "":
            pass  # spacer
        else:
            num += 1
            ws.cell(row=current_row, column=1, value=num).font = bold_font
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).alignment = center_align

            ws.cell(row=current_row, column=2, value=punto).font = data_font
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).alignment = left_align

            ws.cell(row=current_row, column=3, value=revision).font = data_font
            ws.cell(row=current_row, column=3).border = thin_border
            ws.cell(row=current_row, column=3).alignment = left_align

            grav_cell = ws.cell(row=current_row, column=4, value=gravedad)
            grav_cell.border = thin_border
            grav_cell.alignment = center_align
            if gravedad == "GRAVE":
                grav_cell.font = Font(name="Calibri", size=11, bold=True, color=RED_ALERT)
            elif gravedad == "MODERADA":
                grav_cell.font = Font(name="Calibri", size=11, bold=True, color=ORANGE_WARN)
            else:
                grav_cell.font = bold_font

            estado_cell = ws.cell(row=current_row, column=5)
            estado_cell.fill = input_fill
            estado_cell.border = thin_border
            estado_validation.add(estado_cell)

        current_row += 1

    # Summary
    current_row += 1
    ws.cell(row=current_row, column=2, value="RESUMEN:").font = bold_font
    ws.cell(row=current_row + 1, column=2, value="Puntos que cumplen:").font = data_font
    ws.cell(row=current_row + 1, column=3, value=f'=COUNTIF(E5:E{current_row-1},"✓ Cumple")').font = bold_font
    ws.cell(row=current_row + 2, column=2, value="Puntos a mejorar:").font = data_font
    ws.cell(row=current_row + 2, column=3, value=f'=COUNTIF(E5:E{current_row-1},"⚠ Mejorar")').font = Font(name="Calibri", size=11, bold=True, color=ORANGE_WARN)
    ws.cell(row=current_row + 3, column=2, value="Puntos que no cumplen:").font = data_font
    ws.cell(row=current_row + 3, column=3, value=f'=COUNTIF(E5:E{current_row-1},"✗ No cumple")').font = Font(name="Calibri", size=11, bold=True, color=RED_ALERT)

    add_footer(ws, current_row + 5, "Los puntos marcados como GRAVE pueden suponer cierre cautelar del establecimiento.")
    add_footer(ws, current_row + 6, "— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro")

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "15-guia-inspeccion-sanidad.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# MAIN — Generate all templates
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n🛡️  Generando Pack de Plantillas APPCC...\n")

    create_temperaturas_diario()
    create_temperaturas_recepcion()
    create_plan_limpieza()
    create_registro_limpieza()
    create_checklist_recepcion()
    create_trazabilidad()
    create_control_plagas()
    create_matriz_alergenos()
    create_control_aceite()
    create_control_agua()
    create_acciones_correctivas()
    create_analisis_haccp()
    create_higiene_personal()
    create_fichas_alergenos()
    create_guia_inspeccion()
    create_formacion()
    create_protocolo_alerta()

    # Summary
    files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.xlsx')]
    print(f"\n✅ {len(files)} archivos generados en {OUTPUT_DIR}")
    total_size = sum(os.path.getsize(os.path.join(OUTPUT_DIR, f)) for f in files)
    print(f"   Tamaño total: {total_size / 1024:.0f} KB\n")
    for f in sorted(files):
        size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
        print(f"   {f} ({size / 1024:.0f} KB)")
