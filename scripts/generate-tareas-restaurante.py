#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Restaurante Casual.
AI Chef Pro — aichef.pro

Pre-filled recurring task checklists organized by:
- Time of day (apertura, servicio, cierre)
- Work area (cocina, sala, barra, terraza)
- Professional role (chef, manager, sala, barra)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Output ──
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Brand Colors ──
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Zone colors
COCINA_COLOR = "E8F5E9"      # green
SALA_COLOR = "E3F2FD"         # blue
BARRA_COLOR = "FFF3E0"        # orange
TERRAZA_COLOR = "F3E5F5"      # purple
LIMPIEZA_COLOR = "FCE4EC"     # pink
ADMIN_COLOR = "FFF8E1"        # yellow
EVENTO_COLOR = "E0F2F1"       # teal

ZONE_COLORS = {
    "Cocina": COCINA_COLOR,
    "Sala": SALA_COLOR,
    "Barra": BARRA_COLOR,
    "Terraza": TERRAZA_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Evento": EVENTO_COLOR,
    "General": LIGHT_GRAY,
}

# ── Fonts ──
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

# ── Fills ──
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

# ── Borders ──
thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

# ── Alignments ──
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80

    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)

    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    """
    Create a task checklist sheet.

    sections = [
        ("Section Title", "zone", [
            ("Task description", "Responsible", "HH:MM"),
            ...
        ]),
        ...
    ]
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # Column widths
    ws.column_dimensions["A"].width = 5    # checkbox
    ws.column_dimensions["B"].width = 45   # task
    ws.column_dimensions["C"].width = 12   # area/zone
    ws.column_dimensions["D"].width = 18   # responsible
    ws.column_dimensions["E"].width = 12   # hora límite
    ws.column_dimensions["F"].width = 12   # completada
    ws.column_dimensions["G"].width = 15   # firma

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font

    ws.merge_cells("A2:G2")
    ws["A2"].value = "Fecha: ___/___/______    Turno: ☐ Mañana  ☐ Tarde  ☐ Noche    Responsable turno: _________________________"
    ws["A2"].font = subtitle_font

    # Headers
    headers = ["☐", "Tarea", "Zona", "Responsable", "Hora Límite", "Hecha", "Firma"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Validation for "Hecha"
    done_validation = DataValidation(type="list", formula1='"✓,—"', allow_blank=True)
    ws.add_data_validation(done_validation)

    current_row = 5

    for section_title, zone, tasks in sections:
        # Section header
        ws.merge_cells(f"A{current_row}:G{current_row}")
        cell = ws.cell(row=current_row, column=1, value=f"  {section_title}")
        cell.font = section_font
        zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
        section_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).fill = section_fill
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1

        for task_desc, responsible, time_limit in tasks:
            # Checkbox
            ws.cell(row=current_row, column=1, value="☐").font = checkbox_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border

            # Task
            ws.cell(row=current_row, column=2, value=task_desc).font = data_font
            ws.cell(row=current_row, column=2).alignment = left_align
            ws.cell(row=current_row, column=2).border = thin_border

            # Zone
            zone_fill = PatternFill(start_color=ZONE_COLORS.get(zone, LIGHT_GRAY),
                                     end_color=ZONE_COLORS.get(zone, LIGHT_GRAY), fill_type="solid")
            ws.cell(row=current_row, column=3, value=zone).font = Font(name="Calibri", size=10, color="666666")
            ws.cell(row=current_row, column=3).fill = zone_fill
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=3).border = thin_border

            # Responsible
            ws.cell(row=current_row, column=4, value=responsible).font = data_font
            ws.cell(row=current_row, column=4).fill = input_fill
            ws.cell(row=current_row, column=4).alignment = center_align
            ws.cell(row=current_row, column=4).border = thin_border

            # Time limit
            ws.cell(row=current_row, column=5, value=time_limit).font = data_font
            ws.cell(row=current_row, column=5).alignment = center_align
            ws.cell(row=current_row, column=5).border = thin_border

            # Done (validation)
            cell_done = ws.cell(row=current_row, column=6)
            cell_done.fill = input_fill
            cell_done.alignment = center_align
            cell_done.border = thin_border
            done_validation.add(cell_done)

            # Signature
            ws.cell(row=current_row, column=7).fill = input_fill
            ws.cell(row=current_row, column=7).border = thin_border

            current_row += 1

        current_row += 1  # spacer between sections

    # Summary row
    current_row += 1
    ws.merge_cells(f"A{current_row}:C{current_row}")
    ws.cell(row=current_row, column=1, value="Tareas completadas:").font = bold_font
    ws.cell(row=current_row, column=4, value=f'=COUNTIF(F5:F{current_row-2},"✓")').font = bold_font
    ws.cell(row=current_row, column=5, value="de").font = data_font
    total_tasks = sum(len(tasks) for _, _, tasks in sections)
    ws.cell(row=current_row, column=6, value=total_tasks).font = bold_font

    current_row += 2
    ws.cell(row=current_row, column=1, value="Verificado por:").font = bold_font
    ws.merge_cells(f"B{current_row}:D{current_row}")
    ws.cell(row=current_row, column=2).fill = input_fill
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=5, value="Firma:").font = bold_font
    ws.merge_cells(f"F{current_row}:G{current_row}")
    ws.cell(row=current_row, column=6).fill = input_fill
    ws.cell(row=current_row, column=6).border = thin_border

    current_row += 2
    ws.cell(row=current_row, column=1,
            value="— Kit de Tareas Recurrentes · Restaurante Casual · AI Chef Pro · aichef.pro").font = small_font

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "portrait"
    ws.print_area = f"A1:G{current_row}"

    return ws


# ══════════════════════════════════════════════════════════════
# WORKBOOK 1: Apertura y Cierre
# ══════════════════════════════════════════════════════════════
def create_apertura_cierre():
    wb = Workbook()
    add_instructions_sheet(wb, "Tareas de Apertura y Cierre — Restaurante Casual", [
        "Cómo usar estas plantillas",
        "",
        "▸ Imprime la hoja correspondiente al turno (apertura o cierre)",
        "▸ El responsable del turno reparte las tareas al equipo",
        "▸ Cada persona marca ✓ cuando completa su tarea y firma",
        "▸ Al final del turno, el encargado verifica y firma al pie",
        "▸ Archiva las hojas firmadas — demuestran control operativo",
        "",
        "Cómo personalizar",
        "",
        "▸ Las celdas verdes son editables — cambia responsables y horarios",
        "▸ Borra las tareas que no aplican a tu restaurante",
        "▸ Añade tareas específicas de tu local en las filas vacías del final",
        "▸ Ajusta las horas límite a tus horarios reales",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    # ── APERTURA COCINA ──
    create_task_sheet(wb, "Apertura Cocina", "4CAF50",
        "Checklist de Apertura — Cocina",
        [
            ("ENCENDIDO DE EQUIPOS", "Cocina", [
                ("Encender cámaras frigoríficas y verificar temperaturas", "Cocinero apertura", "07:00"),
                ("Encender hornos y precalentar a temperatura de trabajo", "Cocinero apertura", "07:15"),
                ("Encender planchas, freidoras y salamandra", "Cocinero apertura", "07:15"),
                ("Encender campana extractora", "Cocinero apertura", "07:00"),
                ("Verificar gas: llaves abiertas, sin fugas", "Cocinero apertura", "07:00"),
                ("Encender lavavajillas y verificar nivel de producto", "Auxiliar cocina", "07:15"),
            ]),
            ("CONTROL DE TEMPERATURAS", "Cocina", [
                ("Registrar temperatura cámara 1 (refrigeración)", "Cocinero apertura", "07:00"),
                ("Registrar temperatura cámara 2 (refrigeración)", "Cocinero apertura", "07:00"),
                ("Registrar temperatura congelador", "Cocinero apertura", "07:00"),
                ("Registrar en hoja de control de temperaturas APPCC", "Cocinero apertura", "07:15"),
            ]),
            ("REVISIÓN DE MERCANCÍA", "Cocina", [
                ("Revisar caducidades en cámaras (FIFO)", "Sous Chef", "08:00"),
                ("Verificar stock de productos del día", "Sous Chef", "08:00"),
                ("Preparar lista de pedidos urgentes", "Sous Chef", "08:30"),
                ("Recibir e inspeccionar entregas de proveedores", "Sous Chef", "09:00"),
            ]),
            ("MISE EN PLACE", "Cocina", [
                ("Preparar mise en place de partida calientes", "Chef partida", "10:00"),
                ("Preparar mise en place de partida fríos", "Chef partida", "10:00"),
                ("Preparar salsas base del día", "Cocinero", "10:30"),
                ("Cortar vegetales y guarniciones", "Auxiliar cocina", "10:30"),
                ("Preparar postres del día", "Pastelero/Cocinero", "10:30"),
                ("Montar mesa caliente / baño maría", "Cocinero", "11:30"),
                ("Probar platos del día y ajustar sazón", "Jefe de Cocina", "11:45"),
            ]),
            ("LIMPIEZA PRE-SERVICIO", "Cocina", [
                ("Limpiar superficies de trabajo", "Todo el equipo", "11:30"),
                ("Verificar tablas de corte limpias por color", "Sous Chef", "11:30"),
                ("Verificar trapos limpios disponibles", "Auxiliar cocina", "11:30"),
                ("Gel desinfectante en dispensadores", "Auxiliar cocina", "11:30"),
            ]),
        ]
    )

    # ── APERTURA SALA ──
    create_task_sheet(wb, "Apertura Sala", "2196F3",
        "Checklist de Apertura — Sala y Comedor",
        [
            ("MONTAJE DE SALA", "Sala", [
                ("Encender luces y climatización", "Camarero apertura", "11:00"),
                ("Verificar limpieza general de sala", "Camarero apertura", "11:00"),
                ("Montar mesas: manteles/individuales, cubiertos, vasos, servilletas", "Camareros", "11:30"),
                ("Colocar menús/cartas en mesas o atril", "Camarero", "11:30"),
                ("Verificar flores/decoración de mesas", "Camarero", "11:30"),
                ("Preparar estación de servicio: cubiertos extra, servilletas, salsas", "Camarero", "11:30"),
                ("Verificar carta de alérgenos accesible", "Jefe de Sala", "11:30"),
            ]),
            ("MONTAJE DE TERRAZA", "Terraza", [
                ("Sacar mobiliario de terraza", "Camareros", "11:00"),
                ("Limpiar mesas y sillas de terraza", "Camarero", "11:15"),
                ("Montar sombrillas/toldos si aplica", "Camarero", "11:15"),
                ("Verificar calefactores exteriores (invierno)", "Camarero", "11:15"),
                ("Colocar ceniceros y servilleteros", "Camarero", "11:15"),
            ]),
            ("RESERVAS Y BRIEFING", "Sala", [
                ("Revisar libro de reservas del día", "Jefe de Sala", "11:00"),
                ("Asignar mesas a reservas con nombre", "Jefe de Sala", "11:15"),
                ("Verificar eventos especiales (cumpleaños, grupos)", "Jefe de Sala", "11:15"),
                ("Briefing pre-servicio con equipo de sala", "Jefe de Sala", "11:45"),
                ("Comunicar platos del día y alérgenos al equipo", "Jefe de Sala", "11:45"),
                ("Comunicar VIPs y peticiones especiales", "Jefe de Sala", "11:45"),
            ]),
            ("SISTEMAS", "Admin", [
                ("Encender TPV / sistema de cobro", "Encargado", "11:30"),
                ("Verificar rollo de papel en TPV e impresora", "Encargado", "11:30"),
                ("Abrir caja con fondo de cambio", "Encargado", "11:30"),
                ("Verificar datáfono funciona", "Encargado", "11:30"),
                ("Comprobar música ambiental", "Camarero", "11:30"),
            ]),
        ]
    )

    # ── APERTURA BARRA ──
    create_task_sheet(wb, "Apertura Barra", "FF9800",
        "Checklist de Apertura — Barra / Bar",
        [
            ("PREPARACIÓN DE BARRA", "Barra", [
                ("Limpiar barra y superficies", "Barman", "11:00"),
                ("Verificar stock de bebidas (cervezas, vinos, licores)", "Barman", "11:15"),
                ("Reponer neveras de barra", "Barman", "11:15"),
                ("Preparar cubeta de hielo", "Barman", "11:15"),
                ("Cortar garnish: limón, naranja, lima, aceitunas", "Barman", "11:30"),
                ("Preparar zumos frescos del día", "Barman", "11:30"),
                ("Verificar stock de refrescos y agua", "Barman", "11:30"),
            ]),
            ("MÁQUINA DE CAFÉ", "Barra", [
                ("Encender máquina de café y dejar calentar", "Barman", "10:30"),
                ("Purgar grupo de café (2 cafés al vacío)", "Barman", "10:45"),
                ("Verificar stock de leche (entera, desnatada, vegetal)", "Barman", "10:45"),
                ("Limpiar vaporizador de leche", "Barman", "10:45"),
                ("Verificar stock de azúcar, sacarina, estevia", "Barman", "10:45"),
            ]),
            ("CRISTALERÍA Y UTENSILIOS", "Barra", [
                ("Verificar cristalería limpia y sin marcas", "Barman", "11:30"),
                ("Reponer posavasos y servilletas de barra", "Barman", "11:30"),
                ("Verificar coctelera, jigger, colador, muddler", "Barman", "11:30"),
                ("Verificar pajitas y removedores", "Barman", "11:30"),
            ]),
        ]
    )

    # ── CIERRE COCINA ──
    create_task_sheet(wb, "Cierre Cocina", "388E3C",
        "Checklist de Cierre — Cocina",
        [
            ("CONSERVACIÓN DE PRODUCTO", "Cocina", [
                ("Etiquetar y fechar todas las pre-elaboraciones", "Cocineros", "Cierre"),
                ("Guardar mise en place en recipientes herméticos", "Cocineros", "Cierre"),
                ("Aplicar FIFO en cámaras: producto nuevo atrás", "Cocineros", "Cierre"),
                ("Retirar producto caducado o deteriorado", "Sous Chef", "Cierre"),
                ("Cubrir todos los recipientes en cámaras", "Cocineros", "Cierre"),
                ("Registrar temperatura final cámaras y congelador", "Cocinero cierre", "Cierre"),
            ]),
            ("LIMPIEZA DE PARTIDAS", "Cocina", [
                ("Limpiar y desinfectar superficies de trabajo", "Todo el equipo", "Cierre"),
                ("Limpiar plancha, fogones y freidora", "Cocinero partida", "Cierre"),
                ("Limpiar horno y salamandra", "Cocinero partida", "Cierre"),
                ("Vaciar y limpiar baño maría", "Cocinero", "Cierre"),
                ("Limpiar tablas de corte y cuchillos", "Todo el equipo", "Cierre"),
                ("Lavar y guardar utensilios", "Auxiliar cocina", "Cierre"),
            ]),
            ("LIMPIEZA GENERAL", "Limpieza", [
                ("Barrer suelos de cocina", "Auxiliar cocina", "Cierre"),
                ("Fregar suelos de cocina", "Auxiliar cocina", "Cierre"),
                ("Vaciar cubos de basura y colocar bolsas nuevas", "Auxiliar cocina", "Cierre"),
                ("Sacar basura a contenedores", "Auxiliar cocina", "Cierre"),
                ("Limpiar fregaderos", "Auxiliar cocina", "Cierre"),
                ("Verificar trampas de grasa", "Auxiliar cocina", "Cierre"),
            ]),
            ("APAGADO DE EQUIPOS", "Cocina", [
                ("Apagar freidoras y plancha", "Cocinero cierre", "Cierre"),
                ("Apagar hornos y salamandra", "Cocinero cierre", "Cierre"),
                ("Cerrar llaves de gas", "Cocinero cierre", "Cierre"),
                ("Apagar campana extractora", "Cocinero cierre", "Cierre"),
                ("Verificar que cámaras quedan cerradas", "Cocinero cierre", "Cierre"),
                ("Apagar luces de cocina", "Último en salir", "Cierre"),
            ]),
        ]
    )

    # ── CIERRE SALA ──
    create_task_sheet(wb, "Cierre Sala", "1565C0",
        "Checklist de Cierre — Sala, Terraza y Caja",
        [
            ("RECOGIDA DE SALA", "Sala", [
                ("Desmontar mesas: recoger manteles, cubiertos, vasos", "Camareros", "Cierre"),
                ("Limpiar y desinfectar mesas", "Camareros", "Cierre"),
                ("Recoger estación de servicio", "Camarero", "Cierre"),
                ("Reponer servilletas, cubiertos y salsas para mañana", "Camarero", "Cierre"),
                ("Barrer y fregar suelos de sala", "Personal limpieza", "Cierre"),
                ("Verificar baños: limpieza, jabón, papel", "Personal limpieza", "Cierre"),
            ]),
            ("RECOGIDA DE TERRAZA", "Terraza", [
                ("Recoger mesas y sillas de terraza", "Camareros", "Cierre"),
                ("Recoger sombrillas/toldos", "Camarero", "Cierre"),
                ("Barrer terraza", "Camarero", "Cierre"),
                ("Guardar mobiliario si es necesario", "Camarero", "Cierre"),
            ]),
            ("CIERRE DE CAJA", "Admin", [
                ("Imprimir informe Z del TPV", "Encargado", "Cierre"),
                ("Contar efectivo y cuadrar con TPV", "Encargado", "Cierre"),
                ("Registrar propinas si aplica", "Encargado", "Cierre"),
                ("Guardar efectivo en caja fuerte", "Encargado", "Cierre"),
                ("Apagar TPV y datáfono", "Encargado", "Cierre"),
                ("Registrar incidencias del turno", "Encargado", "Cierre"),
            ]),
            ("CIERRE GENERAL", "General", [
                ("Apagar música y TV", "Último en salir", "Cierre"),
                ("Apagar climatización", "Último en salir", "Cierre"),
                ("Verificar ventanas cerradas", "Último en salir", "Cierre"),
                ("Activar alarma", "Último en salir", "Cierre"),
                ("Cerrar puerta con llave", "Último en salir", "Cierre"),
            ]),
        ]
    )

    # ── CIERRE BARRA ──
    create_task_sheet(wb, "Cierre Barra", "E65100",
        "Checklist de Cierre — Barra / Bar",
        [
            ("LIMPIEZA DE BARRA", "Barra", [
                ("Limpiar y desinfectar superficie de barra", "Barman", "Cierre"),
                ("Limpiar máquina de café: purgar, limpiar grupo", "Barman", "Cierre"),
                ("Limpiar vaporizador de leche", "Barman", "Cierre"),
                ("Lavar y guardar cristalería", "Barman", "Cierre"),
                ("Lavar coctelera, jigger y utensilios", "Barman", "Cierre"),
                ("Limpiar fregadero de barra", "Barman", "Cierre"),
            ]),
            ("STOCK Y CONSERVACIÓN", "Barra", [
                ("Guardar garnish sobrante en recipientes", "Barman", "Cierre"),
                ("Verificar stock de cerveza (barriles, botellas)", "Barman", "Cierre"),
                ("Cerrar neveras de barra", "Barman", "Cierre"),
                ("Desechar hielo sobrante y limpiar cubeta", "Barman", "Cierre"),
                ("Anotar productos a pedir para mañana", "Barman", "Cierre"),
                ("Apagar máquina de café", "Barman", "Cierre"),
            ]),
        ]
    )

    path = os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# WORKBOOK 2: Tareas por Partida de Cocina
# ══════════════════════════════════════════════════════════════
def create_partidas_cocina():
    wb = Workbook()
    add_instructions_sheet(wb, "Tareas por Partida de Cocina — Restaurante Casual", [
        "Cómo usar estas plantillas",
        "",
        "▸ Cada pestaña corresponde a una partida o estación de cocina",
        "▸ Las tareas están organizadas: antes, durante y después del servicio",
        "▸ El chef de partida es responsable de completar y firmar",
        "▸ Personaliza añadiendo los platos específicos de tu carta",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    # Partida Calientes
    create_task_sheet(wb, "Calientes", "F44336",
        "Partida de Calientes — Tareas Diarias",
        [
            ("ANTES DEL SERVICIO", "Cocina", [
                ("Encender y precalentar plancha, fogones, horno", "Chef partida", "10:00"),
                ("Preparar fondos y salsas base", "Chef partida", "10:30"),
                ("Mise en place: proteínas porcionadas y temperadas", "Chef partida", "11:00"),
                ("Mise en place: guarniciones calientes listas", "Cocinero", "11:00"),
                ("Verificar stock de aceite para fritura", "Cocinero", "11:00"),
                ("Probar temperatura de aceite de freidora", "Cocinero", "11:30"),
                ("Organizar mesa de pase con platos calientes", "Chef partida", "11:30"),
                ("Probar sazón de todos los platos calientes", "Jefe de Cocina", "11:45"),
            ]),
            ("DURANTE EL SERVICIO", "Cocina", [
                ("Mantener limpieza continua de estación", "Chef partida", "Servicio"),
                ("Control de tiempos de cocción por comanda", "Chef partida", "Servicio"),
                ("Reponer mise en place según se agote", "Cocinero", "Servicio"),
                ("Comunicar 86s (productos agotados) a sala", "Chef partida", "Servicio"),
                ("Control de temperatura de mantenimiento (>65°C)", "Chef partida", "Servicio"),
            ]),
            ("DESPUÉS DEL SERVICIO", "Cocina", [
                ("Guardar proteínas sobrantes: etiquetar y fechar", "Chef partida", "Cierre"),
                ("Guardar salsas: etiquetar y fechar", "Cocinero", "Cierre"),
                ("Limpiar plancha con rasqueta y producto específico", "Chef partida", "Cierre"),
                ("Limpiar fogones y quemadores", "Cocinero", "Cierre"),
                ("Filtrar y/o cambiar aceite de freidora", "Cocinero", "Cierre"),
                ("Limpiar mesa de pase", "Cocinero", "Cierre"),
                ("Dejar partida lista para el siguiente turno", "Chef partida", "Cierre"),
            ]),
        ]
    )

    # Partida Fríos
    create_task_sheet(wb, "Fríos", "2196F3",
        "Partida de Fríos / Ensaladas — Tareas Diarias",
        [
            ("ANTES DEL SERVICIO", "Cocina", [
                ("Lavar y desinfectar lechugas y vegetales", "Cocinero", "10:00"),
                ("Preparar vinagretas y aliños del día", "Cocinero", "10:30"),
                ("Mise en place: cortar vegetales y frutas", "Cocinero", "10:30"),
                ("Preparar bases de ensalada", "Cocinero", "11:00"),
                ("Porcionar carpaccios, ceviches, tartares", "Chef partida", "11:00"),
                ("Verificar stock de productos fríos en cámara", "Chef partida", "11:00"),
                ("Montar estación fría con inserts organizados", "Chef partida", "11:30"),
            ]),
            ("DURANTE EL SERVICIO", "Cocina", [
                ("Montar ensaladas y platos fríos al momento", "Chef partida", "Servicio"),
                ("Mantener temperatura de exposición < 8°C", "Chef partida", "Servicio"),
                ("Reponer inserts de vegetales según se agoten", "Cocinero", "Servicio"),
                ("Limpiar estación entre servicios", "Cocinero", "Servicio"),
            ]),
            ("DESPUÉS DEL SERVICIO", "Cocina", [
                ("Guardar vegetales cortados: tapar y fechar", "Cocinero", "Cierre"),
                ("Guardar vinagretas y aliños en cámara", "Cocinero", "Cierre"),
                ("Desechar hojas marchitas y producto deteriorado", "Cocinero", "Cierre"),
                ("Limpiar y desinfectar estación fría", "Cocinero", "Cierre"),
                ("Limpiar tablas de corte (verde = vegetales)", "Cocinero", "Cierre"),
            ]),
        ]
    )

    # Mise en Place General
    create_task_sheet(wb, "Mise en Place", "4CAF50",
        "Checklist de Mise en Place General",
        [
            ("PROTEÍNAS", "Cocina", [
                ("Carne roja: porcionada, temperada", "Chef partida", "11:00"),
                ("Aves: porcionadas, marinadas si aplica", "Chef partida", "11:00"),
                ("Pescado: limpio, porcionado, en frío", "Chef partida", "11:00"),
                ("Marisco: limpio, desvenado si aplica", "Chef partida", "11:00"),
            ]),
            ("VEGETALES Y GUARNICIONES", "Cocina", [
                ("Patatas: peladas, cortadas, en agua", "Auxiliar", "10:30"),
                ("Cebolla: brunoise, juliana, aros", "Auxiliar", "10:30"),
                ("Ajo: pelado, laminado, picado", "Auxiliar", "10:30"),
                ("Tomate: concassé, rodajas, gajos", "Auxiliar", "10:30"),
                ("Verduras variadas según carta", "Auxiliar", "10:30"),
                ("Hierbas frescas: lavadas, picadas", "Auxiliar", "10:30"),
            ]),
            ("SALSAS Y BASES", "Cocina", [
                ("Fondo oscuro / claro: caliente y listo", "Cocinero", "10:00"),
                ("Salsas del día preparadas", "Cocinero", "11:00"),
                ("Vinagretas y aliños", "Cocinero", "11:00"),
                ("Mayonesas y emulsiones", "Cocinero", "11:00"),
            ]),
            ("POSTRES", "Cocina", [
                ("Postres del día emplatados o listos para emplatar", "Pastelero", "11:30"),
                ("Helados temperados (pasar de -18 a -12°C)", "Pastelero", "11:30"),
                ("Salsas dulces y coulis listos", "Pastelero", "11:30"),
                ("Decoración de postres preparada", "Pastelero", "11:30"),
            ]),
        ]
    )

    path = os.path.join(OUTPUT_DIR, "02-partidas-cocina.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# WORKBOOK 3: Tareas del Manager
# ══════════════════════════════════════════════════════════════
def create_tareas_manager():
    wb = Workbook()
    add_instructions_sheet(wb, "Tareas del Manager / Gerente — Restaurante Casual", [
        "Cómo usar estas plantillas",
        "",
        "▸ Checklist diario: imprime uno por día laborable",
        "▸ Checklist semanal: imprime uno por semana",
        "▸ Checklist mensual: imprime uno por mes",
        "▸ Handover: imprime para cada cambio de turno",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    create_task_sheet(wb, "Diario Manager", "FF9800",
        "Checklist Diario del Manager",
        [
            ("PRIMERA HORA (llegada)", "Admin", [
                ("Revisar email y mensajes pendientes", "Manager", "09:00"),
                ("Revisar reservas del día y asignar mesas", "Manager", "09:00"),
                ("Verificar plantilla del día: ¿alguien ha faltado?", "Manager", "09:00"),
                ("Revisar incidencias del turno anterior", "Manager", "09:15"),
                ("Comprobar entregas de proveedores programadas", "Manager", "09:15"),
            ]),
            ("PRE-SERVICIO", "Admin", [
                ("Recorrido de inspección: cocina, sala, baños, terraza", "Manager", "11:00"),
                ("Verificar limpieza general del local", "Manager", "11:00"),
                ("Hablar con jefe de cocina: platos del día, incidencias", "Manager", "11:15"),
                ("Briefing pre-servicio con equipo de sala", "Manager", "11:45"),
                ("Revisar redes sociales: comentarios, reseñas nuevas", "Manager", "11:30"),
            ]),
            ("DURANTE SERVICIO", "Sala", [
                ("Estar presente en sala durante picos de servicio", "Manager", "Servicio"),
                ("Saludar a clientes habituales / VIPs", "Manager", "Servicio"),
                ("Gestionar reclamaciones o incidencias en sala", "Manager", "Servicio"),
                ("Supervisar tiempos de servicio", "Manager", "Servicio"),
            ]),
            ("POST-SERVICIO", "Admin", [
                ("Revisar comentarios del servicio con jefe cocina", "Manager", "15:30"),
                ("Revisar ventas del mediodía en TPV", "Manager", "15:30"),
                ("Gestionar pedidos a proveedores para mañana", "Manager", "16:00"),
                ("Responder reseñas de Google / TripAdvisor", "Manager", "16:00"),
                ("Actualizar redes sociales si aplica", "Manager", "16:00"),
            ]),
            ("CIERRE DEL DÍA", "Admin", [
                ("Cuadrar caja del día", "Manager", "Cierre"),
                ("Registrar ventas totales", "Manager", "Cierre"),
                ("Anotar incidencias relevantes del día", "Manager", "Cierre"),
                ("Preparar notas para el turno de mañana", "Manager", "Cierre"),
                ("Verificar cierre completo del local", "Manager", "Cierre"),
            ]),
        ]
    )

    create_task_sheet(wb, "Semanal Manager", "E65100",
        "Checklist Semanal del Manager",
        [
            ("LUNES — PLANIFICACIÓN", "Admin", [
                ("Revisar ventas de la semana anterior", "Manager", "Lunes"),
                ("Analizar platos más y menos vendidos", "Manager", "Lunes"),
                ("Planificar carta/menú de la semana", "Manager", "Lunes"),
                ("Revisar y confirmar turnos de la semana", "Manager", "Lunes"),
                ("Reunión semanal con jefe de cocina", "Manager", "Lunes"),
            ]),
            ("MARTES — PROVEEDORES", "Admin", [
                ("Comparar precios de proveedores principales", "Manager", "Martes"),
                ("Gestionar pedidos semanales", "Manager", "Martes"),
                ("Negociar condiciones si hay subidas de precio", "Manager", "Martes"),
                ("Revisar albaranes pendientes de factura", "Manager", "Martes"),
            ]),
            ("MIÉRCOLES — EQUIPO", "Admin", [
                ("Reunión breve con equipo de sala", "Manager", "Miércoles"),
                ("Evaluar rendimiento individual (informal)", "Manager", "Miércoles"),
                ("Gestionar solicitudes de vacaciones/cambios", "Manager", "Miércoles"),
                ("Revisar uniformes y presentación del equipo", "Manager", "Miércoles"),
            ]),
            ("JUEVES — CALIDAD", "Admin", [
                ("Inspección de cámaras: orden, limpieza, FIFO", "Manager", "Jueves"),
                ("Revisión de registros APPCC de la semana", "Manager", "Jueves"),
                ("Verificar stock de productos de limpieza", "Manager", "Jueves"),
                ("Probar 2-3 platos de la carta (control calidad)", "Manager", "Jueves"),
            ]),
            ("VIERNES — MARKETING Y ADMIN", "Admin", [
                ("Planificar publicaciones de redes sociales del fin de semana", "Manager", "Viernes"),
                ("Revisar reservas del fin de semana", "Manager", "Viernes"),
                ("Verificar stock suficiente para fin de semana", "Manager", "Viernes"),
                ("Cerrar semana administrativa: facturas, pagos", "Manager", "Viernes"),
            ]),
        ]
    )

    create_task_sheet(wb, "Mensual Manager", "BF360C",
        "Checklist Mensual del Manager",
        [
            ("FINANZAS", "Admin", [
                ("Revisar P&L del mes (ingresos vs gastos)", "Manager", "Día 1-3"),
                ("Calcular food cost real vs objetivo", "Manager", "Día 1-3"),
                ("Calcular labor cost real vs objetivo", "Manager", "Día 1-3"),
                ("Revisar facturas pendientes de proveedores", "Manager", "Día 1-3"),
                ("Analizar ticket medio y evolución", "Manager", "Día 1-3"),
            ]),
            ("EQUIPO", "Admin", [
                ("Evaluación formal de rendimiento del equipo", "Manager", "Mensual"),
                ("Detectar necesidades de formación", "Manager", "Mensual"),
                ("Revisar cumplimiento de horarios del mes", "Manager", "Mensual"),
                ("Planificar turnos del mes siguiente", "Manager", "Mensual"),
                ("Gestionar nóminas y variables", "Manager", "Mensual"),
            ]),
            ("MANTENIMIENTO", "Admin", [
                ("Revisión de equipos de cocina (estado general)", "Manager", "Mensual"),
                ("Programar mantenimiento preventivo si necesario", "Manager", "Mensual"),
                ("Revisar extintores y salidas de emergencia", "Manager", "Mensual"),
                ("Verificar estado de mobiliario (mesas, sillas)", "Manager", "Mensual"),
            ]),
            ("MARKETING Y ESTRATEGIA", "Admin", [
                ("Analizar reseñas del mes: tendencias y puntos débiles", "Manager", "Mensual"),
                ("Planificar promociones o eventos del mes siguiente", "Manager", "Mensual"),
                ("Actualizar carta si hay cambios estacionales", "Manager", "Mensual"),
                ("Revisar posicionamiento en Google Maps", "Manager", "Mensual"),
            ]),
        ]
    )

    create_task_sheet(wb, "Handover Turno", "9C27B0",
        "Checklist de Cambio de Turno (Handover)",
        [
            ("TRASPASO DE INFORMACIÓN", "General", [
                ("Incidencias del turno que sale", "Encargado saliente", "Cambio"),
                ("Mesas ocupadas y estado de cada una", "Encargado saliente", "Cambio"),
                ("Reservas pendientes del turno entrante", "Encargado saliente", "Cambio"),
                ("Productos agotados (86s)", "Encargado saliente", "Cambio"),
                ("Reclamaciones o clientes especiales", "Encargado saliente", "Cambio"),
            ]),
            ("ESTADO DEL LOCAL", "General", [
                ("Estado de limpieza: cocina, sala, baños", "Encargado saliente", "Cambio"),
                ("Estado de caja: fondo correcto", "Encargado saliente", "Cambio"),
                ("Pedidos pendientes de recibir", "Encargado saliente", "Cambio"),
                ("Tareas pendientes no completadas", "Encargado saliente", "Cambio"),
            ]),
            ("CONFIRMACIÓN TURNO ENTRANTE", "General", [
                ("Turno entrante confirma recepción de información", "Encargado entrante", "Cambio"),
                ("Recorrido rápido de verificación", "Encargado entrante", "Cambio"),
                ("Firma de traspaso: ambos encargados", "Ambos", "Cambio"),
            ]),
        ]
    )

    path = os.path.join(OUTPUT_DIR, "03-tareas-manager.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# WORKBOOK 4: Tareas por Perfil Profesional
# ══════════════════════════════════════════════════════════════
def create_tareas_perfiles():
    wb = Workbook()
    add_instructions_sheet(wb, "Tareas por Perfil Profesional — Restaurante Casual", [
        "Cómo usar estas plantillas",
        "",
        "▸ Cada pestaña es un perfil profesional diferente",
        "▸ Entrega la hoja correspondiente a cada miembro del equipo",
        "▸ Son sus responsabilidades diarias — no negociables",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    create_task_sheet(wb, "Jefe de Cocina", "4CAF50",
        "Tareas Diarias — Jefe de Cocina / Chef Ejecutivo",
        [
            ("MAÑANA", "Cocina", [
                ("Verificar que todo el equipo de cocina ha llegado", "Jefe Cocina", "08:00"),
                ("Revisar mise en place de todas las partidas", "Jefe Cocina", "11:00"),
                ("Probar sazón de todas las elaboraciones", "Jefe Cocina", "11:45"),
                ("Aprobar presentación de platos del día", "Jefe Cocina", "11:45"),
                ("Comunicar platos del día y 86s a sala", "Jefe Cocina", "11:45"),
                ("Supervisar recepción de mercancía", "Jefe Cocina", "Según entrega"),
            ]),
            ("SERVICIO", "Cocina", [
                ("Dirigir el pase: control de tiempos y calidad", "Jefe Cocina", "Servicio"),
                ("Verificar presentación de cada plato antes de salir", "Jefe Cocina", "Servicio"),
                ("Coordinar entre partidas", "Jefe Cocina", "Servicio"),
                ("Gestionar devoluciones o reclamaciones de cocina", "Jefe Cocina", "Servicio"),
            ]),
            ("POST-SERVICIO", "Cocina", [
                ("Evaluar servicio con sous chef: qué mejorar", "Jefe Cocina", "15:30"),
                ("Supervisar cierre de todas las partidas", "Jefe Cocina", "Cierre"),
                ("Planificar carta/especiales del día siguiente", "Jefe Cocina", "16:00"),
                ("Revisar pedidos a proveedores", "Jefe Cocina", "16:00"),
                ("Actualizar fichas técnicas si hay cambios", "Jefe Cocina", "Semanal"),
            ]),
        ]
    )

    create_task_sheet(wb, "Sous Chef", "388E3C",
        "Tareas Diarias — Sous Chef",
        [
            ("MAÑANA", "Cocina", [
                ("Organizar equipo de cocina al llegar", "Sous Chef", "08:00"),
                ("Supervisar FIFO en cámaras", "Sous Chef", "08:30"),
                ("Preparar lista de pedidos urgentes", "Sous Chef", "09:00"),
                ("Supervisar mise en place de todas las partidas", "Sous Chef", "10:00-11:30"),
                ("Cubrir partida si falta personal", "Sous Chef", "Si aplica"),
            ]),
            ("SERVICIO", "Cocina", [
                ("Apoyar en la partida más cargada", "Sous Chef", "Servicio"),
                ("Cubrir al jefe de cocina en su ausencia", "Sous Chef", "Si aplica"),
                ("Control de calidad de platos en pase", "Sous Chef", "Servicio"),
            ]),
            ("POST-SERVICIO", "Cocina", [
                ("Supervisar limpieza y cierre de cocina", "Sous Chef", "Cierre"),
                ("Inventario rápido de productos críticos", "Sous Chef", "Cierre"),
                ("Preparar briefing para equipo del día siguiente", "Sous Chef", "Cierre"),
            ]),
        ]
    )

    create_task_sheet(wb, "Jefe de Sala", "2196F3",
        "Tareas Diarias — Jefe de Sala / Maître",
        [
            ("MAÑANA", "Sala", [
                ("Revisar libro de reservas del día", "Jefe Sala", "10:00"),
                ("Asignar mesas y secciones a camareros", "Jefe Sala", "10:30"),
                ("Verificar montaje completo de sala y terraza", "Jefe Sala", "11:30"),
                ("Briefing pre-servicio: platos del día, alérgenos, VIPs", "Jefe Sala", "11:45"),
                ("Verificar presentación del equipo (uniforme)", "Jefe Sala", "11:45"),
            ]),
            ("SERVICIO", "Sala", [
                ("Recibir y acomodar clientes", "Jefe Sala", "Servicio"),
                ("Supervisar tiempos de mesa", "Jefe Sala", "Servicio"),
                ("Gestionar reclamaciones en sala", "Jefe Sala", "Servicio"),
                ("Coordinar con cocina: tiempos, especiales, alérgenos", "Jefe Sala", "Servicio"),
                ("Gestionar walk-ins y lista de espera", "Jefe Sala", "Servicio"),
            ]),
            ("POST-SERVICIO", "Sala", [
                ("Revisar reseñas y comentarios del día", "Jefe Sala", "15:30"),
                ("Evaluar servicio con equipo de sala", "Jefe Sala", "15:30"),
                ("Revisar reservas de mañana", "Jefe Sala", "Cierre"),
                ("Supervisar recogida y cierre de sala", "Jefe Sala", "Cierre"),
            ]),
        ]
    )

    create_task_sheet(wb, "Camarero", "1565C0",
        "Tareas Diarias — Camarero",
        [
            ("PRE-SERVICIO", "Sala", [
                ("Montar mesas de tu sección", "Camarero", "11:00"),
                ("Verificar limpieza de tu sección", "Camarero", "11:15"),
                ("Preparar estación de servicio", "Camarero", "11:15"),
                ("Repasar carta y platos del día", "Camarero", "11:30"),
                ("Asistir al briefing pre-servicio", "Camarero", "11:45"),
            ]),
            ("SERVICIO", "Sala", [
                ("Dar la bienvenida al cliente con sonrisa", "Camarero", "Servicio"),
                ("Ofrecer carta de bebidas inmediatamente", "Camarero", "Servicio"),
                ("Tomar comanda completa (sin volver 2 veces)", "Camarero", "Servicio"),
                ("Preguntar por alérgenos antes de tomar comanda", "Camarero", "Servicio"),
                ("Servir bebidas en < 3 minutos", "Camarero", "Servicio"),
                ("Retirar platos cuando TODOS los comensales terminen", "Camarero", "Servicio"),
                ("Ofrecer postre y café al finalizar", "Camarero", "Servicio"),
                ("Presentar cuenta cuando la pidan (nunca antes)", "Camarero", "Servicio"),
            ]),
            ("POST-SERVICIO", "Sala", [
                ("Desmontar y limpiar mesas de tu sección", "Camarero", "Cierre"),
                ("Reponer estación de servicio", "Camarero", "Cierre"),
                ("Limpiar menús y cartas", "Camarero", "Cierre"),
            ]),
        ]
    )

    create_task_sheet(wb, "Barman", "FF9800",
        "Tareas Diarias — Barman / Bartender",
        [
            ("PRE-SERVICIO", "Barra", [
                ("Abrir barra: limpieza, stock, hielo, garnish", "Barman", "10:30"),
                ("Preparar cafés de prueba (calibrar molinillo)", "Barman", "10:45"),
                ("Revisar vinos por copa: ¿botellas abiertas OK?", "Barman", "11:00"),
                ("Montar mise en place de cocktails si aplica", "Barman", "11:15"),
            ]),
            ("SERVICIO", "Barra", [
                ("Preparar bebidas en < 2 minutos", "Barman", "Servicio"),
                ("Mantener barra limpia y organizada siempre", "Barman", "Servicio"),
                ("Reponer hielo, refrescos y cervezas continuamente", "Barman", "Servicio"),
                ("Atender clientes de barra con prioridad", "Barman", "Servicio"),
            ]),
            ("POST-SERVICIO", "Barra", [
                ("Limpiar barra, fregadero y máquina de café", "Barman", "Cierre"),
                ("Guardar garnish y productos perecederos", "Barman", "Cierre"),
                ("Hacer inventario rápido de cerveza y licores", "Barman", "Cierre"),
                ("Anotar pedidos necesarios para mañana", "Barman", "Cierre"),
            ]),
        ]
    )

    path = os.path.join(OUTPUT_DIR, "04-tareas-perfiles.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# WORKBOOK 5: Tareas Semanales y Mensuales
# ══════════════════════════════════════════════════════════════
def create_tareas_periodicas():
    wb = Workbook()
    add_instructions_sheet(wb, "Tareas Semanales y Mensuales — Restaurante Casual", [
        "Cómo usar estas plantillas",
        "",
        "▸ Imprime la hoja semanal cada lunes y la mensual cada día 1",
        "▸ Asigna cada tarea a un responsable con fecha límite",
        "▸ Archiva las hojas firmadas junto con los registros APPCC",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    create_task_sheet(wb, "Limpieza Semanal", "4CAF50",
        "Limpieza Profunda Semanal",
        [
            ("COCINA — Limpieza profunda", "Limpieza", [
                ("Limpiar interior de cámaras frigoríficas", "Equipo cocina", "Lunes"),
                ("Desmontar y limpiar quemadores de fogones", "Cocinero", "Martes"),
                ("Limpiar horno por dentro (pirolisis o manual)", "Cocinero", "Martes"),
                ("Desmontar y limpiar filtros de campana", "Auxiliar", "Miércoles"),
                ("Limpiar freidora: vaciar, desengrasante, aclarar", "Cocinero", "Miércoles"),
                ("Limpiar interior de microondas y salamandra", "Auxiliar", "Jueves"),
                ("Limpiar rejillas de ventilación", "Auxiliar", "Jueves"),
                ("Limpiar detrás de equipos (mover si es posible)", "Equipo cocina", "Viernes"),
                ("Descalcificar lavavajillas", "Auxiliar", "Viernes"),
            ]),
            ("SALA Y BARRA", "Limpieza", [
                ("Limpiar cristales y espejos", "Personal limpieza", "Lunes"),
                ("Limpiar sillas por debajo y patas", "Camarero", "Martes"),
                ("Limpiar vitrina de barra por dentro", "Barman", "Miércoles"),
                ("Descalcificar máquina de café", "Barman", "Miércoles"),
                ("Limpiar conductos de cerveza de barril", "Barman", "Viernes"),
            ]),
            ("BAÑOS Y GENERAL", "Limpieza", [
                ("Limpieza profunda de baños (azulejos, juntas)", "Personal limpieza", "Jueves"),
                ("Limpiar vestuarios del personal", "Personal limpieza", "Jueves"),
                ("Limpiar terraza a fondo (suelos, macetas)", "Camarero", "Sábado"),
                ("Revisar y reponer productos de limpieza", "Encargado", "Viernes"),
            ]),
        ]
    )

    create_task_sheet(wb, "FIFO Semanal", "FF9800",
        "Revisión FIFO Semanal — Control de Caducidades",
        [
            ("CÁMARA REFRIGERACIÓN 1", "Cocina", [
                ("Verificar etiquetado y fechas de todas las bandejas", "Sous Chef", "Lunes"),
                ("Reorganizar: producto más antiguo delante", "Cocinero", "Lunes"),
                ("Retirar producto caducado o sin etiquetar", "Cocinero", "Lunes"),
                ("Verificar que crudo y cocinado están separados", "Sous Chef", "Lunes"),
            ]),
            ("CÁMARA REFRIGERACIÓN 2", "Cocina", [
                ("Verificar etiquetado y fechas", "Sous Chef", "Lunes"),
                ("Reorganizar FIFO", "Cocinero", "Lunes"),
                ("Retirar producto caducado", "Cocinero", "Lunes"),
            ]),
            ("CONGELADOR", "Cocina", [
                ("Verificar etiquetado y fechas de congelación", "Sous Chef", "Lunes"),
                ("Comprobar que no hay producto > 3 meses congelado", "Sous Chef", "Lunes"),
                ("Verificar ausencia de quemaduras por frío", "Sous Chef", "Lunes"),
                ("Reorganizar por categoría", "Cocinero", "Lunes"),
            ]),
            ("ALMACÉN SECO", "Cocina", [
                ("Verificar caducidades de conservas y secos", "Auxiliar", "Martes"),
                ("Reorganizar por fecha: más antiguo delante", "Auxiliar", "Martes"),
                ("Verificar ausencia de plagas (polillas, hormigas)", "Auxiliar", "Martes"),
                ("Comprobar envases cerrados y sin humedad", "Auxiliar", "Martes"),
            ]),
        ]
    )

    create_task_sheet(wb, "Mantenimiento Mensual", "795548",
        "Mantenimiento Mensual de Equipos",
        [
            ("EQUIPOS DE COCINA", "Cocina", [
                ("Revisión general de horno: resistencias, puerta, juntas", "Técnico/Chef", "Mensual"),
                ("Verificar calibración de termostatos", "Técnico/Chef", "Mensual"),
                ("Revisión de campana extractora: motor, filtros, conducto", "Empresa externa", "Mensual"),
                ("Verificar estado de quemadores y pilotos", "Chef", "Mensual"),
                ("Limpiar condensador de cámaras frigoríficas", "Técnico", "Mensual"),
                ("Verificar juntas de puertas de cámaras", "Chef", "Mensual"),
            ]),
            ("SALA Y BARRA", "Sala", [
                ("Verificar estado del mobiliario: mesas, sillas, bancos", "Manager", "Mensual"),
                ("Revisar iluminación: bombillas fundidas, focos", "Manager", "Mensual"),
                ("Revisión de climatización: filtros, temperatura", "Técnico", "Mensual"),
                ("Verificar estado de TPV y datáfono", "Manager", "Mensual"),
            ]),
            ("SEGURIDAD", "General", [
                ("Verificar extintores: carga, accesibilidad, caducidad", "Manager", "Mensual"),
                ("Comprobar salidas de emergencia: señalización, acceso libre", "Manager", "Mensual"),
                ("Verificar botiquín: stock completo, caducidades", "Manager", "Mensual"),
                ("Comprobar detectores de humo", "Manager", "Mensual"),
            ]),
        ]
    )

    path = os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# WORKBOOK 6: Eventos y Festivos
# ══════════════════════════════════════════════════════════════
def create_tareas_eventos():
    wb = Workbook()
    add_instructions_sheet(wb, "Tareas para Eventos y Festivos — Restaurante Casual", [
        "Cómo usar estas plantillas",
        "",
        "▸ Usa la pestaña Pre-Evento para planificar cualquier evento",
        "▸ Las pestañas de festivos son para fechas señaladas del año",
        "▸ Personaliza con las tradiciones y menú de tu restaurante",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    create_task_sheet(wb, "Pre-Evento", "E91E63",
        "Checklist Pre-Evento (48h → Día del Evento)",
        [
            ("48 HORAS ANTES", "Admin", [
                ("Confirmar número definitivo de comensales", "Manager", "48h antes"),
                ("Confirmar menú con cliente (si es privado)", "Manager", "48h antes"),
                ("Verificar stock de ingredientes para el menú", "Jefe Cocina", "48h antes"),
                ("Hacer pedidos extra a proveedores", "Sous Chef", "48h antes"),
                ("Confirmar personal extra si es necesario", "Manager", "48h antes"),
                ("Preparar decoración especial si aplica", "Manager", "48h antes"),
            ]),
            ("24 HORAS ANTES", "Admin", [
                ("Recibir y verificar pedidos extra", "Sous Chef", "24h antes"),
                ("Iniciar pre-elaboraciones que lo permitan", "Equipo cocina", "24h antes"),
                ("Confirmar asignación de mesas y disposición sala", "Jefe Sala", "24h antes"),
                ("Briefing especial con todo el equipo", "Manager", "24h antes"),
                ("Verificar vajilla y cristalería suficiente", "Jefe Sala", "24h antes"),
            ]),
            ("DÍA DEL EVENTO — MAÑANA", "Evento", [
                ("Montaje especial de sala según plano", "Camareros", "Mañana"),
                ("Verificar decoración, flores, velas si aplica", "Jefe Sala", "Mañana"),
                ("Mise en place completa para el evento", "Equipo cocina", "Mañana"),
                ("Test de sonido si hay música/discurso", "Manager", "Mañana"),
                ("Briefing final 30 min antes del evento", "Manager", "30 min antes"),
            ]),
            ("DÍA DEL EVENTO — POST", "Evento", [
                ("Desmontaje de sala al terminar", "Camareros", "Post-evento"),
                ("Limpieza profunda extra", "Todo el equipo", "Post-evento"),
                ("Evaluación del evento con equipo", "Manager", "Post-evento"),
                ("Enviar agradecimiento al cliente (email/mensaje)", "Manager", "Día siguiente"),
            ]),
        ]
    )

    create_task_sheet(wb, "San Valentín", "E91E63",
        "Tareas Específicas — San Valentín (14 de Febrero)",
        [
            ("1 SEMANA ANTES", "Admin", [
                ("Diseñar menú especial de San Valentín", "Jefe Cocina", "7 días antes"),
                ("Publicar menú en redes sociales y web", "Manager", "7 días antes"),
                ("Abrir reservas con sistema de prepago o confirmación", "Manager", "7 días antes"),
                ("Comprar decoración temática (flores, pétalos, velas)", "Manager", "5 días antes"),
                ("Planificar turnos reforzados", "Manager", "5 días antes"),
            ]),
            ("DÍA DE SAN VALENTÍN", "Evento", [
                ("Decorar sala con ambientación romántica", "Camareros", "Mañana"),
                ("Preparar detalle de bienvenida por mesa (rosa, bombón)", "Jefe Sala", "Mañana"),
                ("Preparar cocktail de bienvenida si está en menú", "Barman", "Pre-servicio"),
                ("Servicio especial: ritmo pausado, atención premium", "Todo equipo sala", "Servicio"),
                ("Ofrecer postre con presentación especial", "Pastelero", "Servicio"),
            ]),
        ]
    )

    create_task_sheet(wb, "Navidad", "4CAF50",
        "Tareas Específicas — Navidad y Nochevieja",
        [
            ("2 SEMANAS ANTES", "Admin", [
                ("Diseñar menú navideño / menú de Nochevieja", "Jefe Cocina", "15 días antes"),
                ("Publicar menú y abrir reservas", "Manager", "15 días antes"),
                ("Hacer pedidos especiales (marisco, cordero, turrón)", "Sous Chef", "10 días antes"),
                ("Planificar turnos de todo el periodo festivo", "Manager", "10 días antes"),
                ("Comprar decoración navideña", "Manager", "10 días antes"),
            ]),
            ("SEMANA DE NAVIDAD", "Evento", [
                ("Montar decoración navideña en sala y barra", "Camareros", "Lunes"),
                ("Verificar stock para toda la semana (sin entregas)", "Sous Chef", "Lunes"),
                ("Pre-elaboraciones extra para los días de más volumen", "Equipo cocina", "Martes-Miérc"),
                ("Briefing especial: servicio premium, sin prisas", "Manager", "Cada día"),
                ("Stock extra de cava/champagne para Nochevieja", "Barman", "30 diciembre"),
            ]),
            ("NOCHEVIEJA", "Evento", [
                ("Mise en place completa antes de las 18:00", "Equipo cocina", "18:00"),
                ("Decorar para la noche: cotillón, uvas, elementos especiales", "Camareros", "19:00"),
                ("Preparar uvas y cava para las campanadas", "Barman", "23:30"),
                ("Servir copa de cava a todos los clientes a medianoche", "Todo equipo", "00:00"),
            ]),
        ]
    )

    create_task_sheet(wb, "Temporada Terraza", "FF9800",
        "Tareas de Inicio / Fin de Temporada de Terraza",
        [
            ("APERTURA DE TEMPORADA", "Terraza", [
                ("Limpiar a fondo todo el mobiliario de terraza", "Camareros", "Día apertura"),
                ("Verificar estado de mesas y sillas: reparar/sustituir", "Manager", "1 semana antes"),
                ("Verificar toldo/sombrilla: estado y funcionamiento", "Manager", "1 semana antes"),
                ("Solicitar licencia de terraza al Ayuntamiento si aplica", "Manager", "1 mes antes"),
                ("Preparar carta de terraza si es diferente", "Jefe Cocina", "1 semana antes"),
                ("Verificar iluminación exterior", "Manager", "Día apertura"),
            ]),
            ("CIERRE DE TEMPORADA", "Terraza", [
                ("Limpiar a fondo mobiliario antes de guardar", "Camareros", "Último día"),
                ("Guardar mobiliario en almacén", "Camareros", "Último día"),
                ("Guardar sombrillas/toldos protegidos", "Camareros", "Último día"),
                ("Desmontar calefactores exteriores si aplica", "Técnico", "Último día"),
                ("Inventario de mobiliario: anotar roturas/reposiciones", "Manager", "Último día"),
            ]),
        ]
    )

    path = os.path.join(OUTPUT_DIR, "06-eventos-festivos.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# WORKBOOK 7: Plantilla Maestra Personalizable
# ══════════════════════════════════════════════════════════════
def create_plantilla_personalizable():
    wb = Workbook()
    add_instructions_sheet(wb, "Plantilla Maestra Personalizable", [
        "Cómo usar esta plantilla",
        "",
        "▸ Usa estas hojas en blanco para crear tus propias listas de tareas",
        "▸ Copia la estructura de las plantillas pre-rellenadas",
        "▸ Personaliza: añade tareas, cambia responsables, ajusta horarios",
        "▸ Puedes duplicar pestañas (click derecho → 'Mover o copiar')",
        "",
        "Consejos de personalización",
        "",
        "▸ Agrupa las tareas por franja horaria (apertura, servicio, cierre)",
        "▸ Asigna siempre un responsable — las tareas sin dueño no se hacen",
        "▸ Pon hora límite realista — mejor ser específico que genérico",
        "▸ Revisa y actualiza las listas cada mes con tu equipo",
        "▸ Imprime en A4 y deja espacio para firma del encargado",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    for sheet_name, title in [
        ("Por Franja Horaria", "Plantilla en Blanco — Por Franja Horaria"),
        ("Por Área", "Plantilla en Blanco — Por Área de Trabajo"),
        ("Por Perfil", "Plantilla en Blanco — Por Perfil Profesional"),
    ]:
        create_task_sheet(wb, sheet_name, GOLD, title, [
            ("(Sección 1 — Personaliza este título)", "General", [
                ("(Escribe aquí tu tarea)", "(Responsable)", "(Hora)"),
                ("", "", ""),
                ("", "", ""),
                ("", "", ""),
                ("", "", ""),
            ]),
            ("(Sección 2 — Personaliza este título)", "General", [
                ("(Escribe aquí tu tarea)", "(Responsable)", "(Hora)"),
                ("", "", ""),
                ("", "", ""),
                ("", "", ""),
                ("", "", ""),
            ]),
            ("(Sección 3 — Personaliza este título)", "General", [
                ("(Escribe aquí tu tarea)", "(Responsable)", "(Hora)"),
                ("", "", ""),
                ("", "", ""),
                ("", "", ""),
                ("", "", ""),
            ]),
        ])

    path = os.path.join(OUTPUT_DIR, "07-plantilla-personalizable.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# BONUS 1: Briefing de Servicio
# ══════════════════════════════════════════════════════════════
def create_briefing_servicio():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS: Plantilla de Briefing Pre-Servicio", [
        "Cómo usar esta plantilla",
        "",
        "▸ El jefe de sala o manager rellena antes del briefing",
        "▸ Se comunica al equipo en la reunión de 5-10 minutos pre-servicio",
        "▸ Imprime y pega en la zona de pase para referencia durante servicio",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Briefing")
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50

    ws.merge_cells("B1:C1")
    ws["B1"].value = "Briefing Pre-Servicio"
    ws["B1"].font = title_font

    ws["B2"].value = "Fecha: ___/___/______    Turno: ☐ Comida  ☐ Cena    Encargado: _________________________"
    ws["B2"].font = subtitle_font

    sections = [
        ("RESERVAS Y OCUPACIÓN", [
            ("Nº de reservas:", ""),
            ("Nº de comensales previstos:", ""),
            ("Mesas VIP / clientes especiales:", ""),
            ("Grupos o eventos:", ""),
            ("Ocupación estimada:", "_____ %"),
        ]),
        ("PLATOS DEL DÍA Y 86s", [
            ("Plato del día 1:", ""),
            ("Plato del día 2:", ""),
            ("Postre especial:", ""),
            ("Productos AGOTADOS (86):", ""),
            ("Cambios en carta:", ""),
        ]),
        ("ALÉRGENOS Y ESPECIALES", [
            ("Reservas con alérgenos declarados:", ""),
            ("Dietas especiales (vegano, celíaco...):", ""),
            ("Restricciones a comunicar:", ""),
        ]),
        ("EQUIPO DEL TURNO", [
            ("Personal cocina:", ""),
            ("Personal sala:", ""),
            ("Personal barra:", ""),
            ("¿Falta alguien? Cobertura:", ""),
        ]),
        ("NOTAS E INCIDENCIAS", [
            ("Incidencias del turno anterior:", ""),
            ("Objetivos del turno:", ""),
            ("Mensaje del manager:", ""),
        ]),
    ]

    current_row = 4
    for section_title, items in sections:
        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws.cell(row=current_row, column=2, value=section_title).font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        ws.cell(row=current_row, column=2).fill = header_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).fill = header_fill
        ws.cell(row=current_row, column=3).border = thin_border
        current_row += 1

        for label, default in items:
            ws.cell(row=current_row, column=2, value=label).font = bold_font
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=3, value=default).font = data_font
            ws.cell(row=current_row, column=3).fill = input_fill
            ws.cell(row=current_row, column=3).border = thin_border
            current_row += 1

        current_row += 1

    ws.cell(row=current_row + 1, column=2, value="Firma encargado: _________________________").font = subtitle_font
    ws.cell(row=current_row + 3, column=2,
            value="— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro").font = small_font

    ws.page_setup.orientation = "portrait"

    path = os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# BONUS 2: Calendario Anual de Tareas Especiales
# ══════════════════════════════════════════════════════════════
def create_calendario_anual():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS: Calendario Anual de Tareas Especiales", [
        "Cómo usar esta plantilla",
        "",
        "▸ Consulta las fechas clave de hostelería mes a mes",
        "▸ Planifica con antelación: menús especiales, decoración, stock",
        "▸ Añade las fechas relevantes de tu zona (fiestas locales, ferias)",
        "",
        "— Kit de Tareas Recurrentes · AI Chef Pro · aichef.pro",
    ])

    ws = wb.create_sheet(title="Calendario Anual")
    ws.sheet_properties.tabColor = GOLD

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Calendario Anual — Fechas Clave de Hostelería"
    ws["A1"].font = title_font

    ws["A2"].value = "Año: ______    Añade tus fechas locales en las filas vacías"
    ws["A2"].font = subtitle_font

    headers = ["Mes", "Fecha / Evento", "Tareas Clave", "Antelación"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    eventos = [
        ("Enero", "1 Ene — Año Nuevo", "Menú brunch recuperación, stock reducido", "1 semana"),
        ("Enero", "6 Ene — Reyes", "Roscón de Reyes, menú familiar", "1 semana"),
        ("Febrero", "14 Feb — San Valentín", "Menú parejas, decoración romántica, turnos extra", "2 semanas"),
        ("Febrero", "Carnaval (variable)", "Menú temático, decoración, disfraces equipo", "1 semana"),
        ("Marzo-Abril", "Semana Santa (variable)", "Menú de vigilia, bacalao, potaje, torrijas", "2 semanas"),
        ("Marzo-Abril", "Apertura terraza", "Montaje, licencia, carta exterior", "1 mes"),
        ("Mayo", "1 Mayo — Día del Trabajo", "Servicio de festivo, turnos especiales", "1 semana"),
        ("Mayo", "Primer domingo — Día de la Madre", "Menú especial, reservas ampliadas, detalle", "2 semanas"),
        ("Junio", "Inicio temporada alta", "Refuerzo de plantilla, stock verano", "1 mes"),
        ("Junio", "24 Jun — San Juan", "Menú especial, terraza nocturna", "1 semana"),
        ("Julio-Agosto", "Temporada alta", "Personal extra, horarios ampliados, carta verano", "Continuo"),
        ("Septiembre", "Vuelta al cole / rutina", "Carta otoño, menú del día reforzado", "2 semanas"),
        ("Octubre", "31 Oct — Halloween", "Decoración temática, menú/cocktails especiales", "1 semana"),
        ("Noviembre", "Cierre terraza", "Recoger mobiliario, inventario, reparaciones", "1 semana"),
        ("Diciembre", "Cenas de empresa", "Menú de grupo, reservas grandes, turnos extra", "1 mes"),
        ("Diciembre", "24-25 Dic — Nochebuena/Navidad", "Menú navideño, decoración, stock especial", "2 semanas"),
        ("Diciembre", "31 Dic — Nochevieja", "Menú gala, cotillón, uvas, cava, turnos noche", "2 semanas"),
        ("", "", "", ""),
        ("(Tu fecha)", "(Añade aquí)", "(Tareas específicas)", ""),
        ("(Tu fecha)", "(Añade aquí)", "(Tareas específicas)", ""),
        ("(Tu fecha)", "(Añade aquí)", "(Tareas específicas)", ""),
        ("(Tu fecha)", "(Añade aquí)", "(Tareas específicas)", ""),
        ("(Tu fecha)", "(Añade aquí)", "(Tareas específicas)", ""),
    ]

    for i, (mes, evento, tareas, antelacion) in enumerate(eventos):
        row = 5 + i
        fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if i % 2 == 0 else PatternFill()

        if mes.startswith("("):
            fill = input_fill

        values = [mes, evento, tareas, antelacion]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = left_align if col_idx in (2, 3) else center_align

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"

    path = os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n📋 Generando Kit de Tareas Recurrentes — Restaurante Casual...\n")

    create_apertura_cierre()
    create_partidas_cocina()
    create_tareas_manager()
    create_tareas_perfiles()
    create_tareas_periodicas()
    create_tareas_eventos()
    create_plantilla_personalizable()
    create_briefing_servicio()
    create_calendario_anual()

    files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.xlsx')]
    print(f"\n✅ {len(files)} archivos generados en {OUTPUT_DIR}")
    total_size = sum(os.path.getsize(os.path.join(OUTPUT_DIR, f)) for f in files)
    print(f"   Tamaño total: {total_size / 1024:.0f} KB\n")
    for f in sorted(files):
        size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
        print(f"   {f} ({size / 1024:.0f} KB)")
