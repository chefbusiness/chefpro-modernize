#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Bar / Cocktails.
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-bar"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

BARRA_COLOR = "FFF3E0"
COCTELERIA_COLOR = "E3F2FD"
BODEGA_COLOR = "F3E5F5"
TERRAZA_COLOR = "E8F5E9"
LIMPIEZA_COLOR = "FCE4EC"
ADMIN_COLOR = "FFF8E1"
EVENTO_COLOR = "E0F2F1"
CAFE_COLOR = "EFEBE9"

ZONE_COLORS = {
    "Barra": BARRA_COLOR,
    "Coctelería": COCTELERIA_COLOR,
    "Bodega": BODEGA_COLOR,
    "Terraza": TERRAZA_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Evento": EVENTO_COLOR,
    "Café": CAFE_COLOR,
    "General": LIGHT_GRAY,
}

title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"].value = "Fecha: ___/___/______    Turno: ☐ Mañana  ☐ Tarde  ☐ Noche    Responsable turno: _________________________"
    ws["A2"].font = subtitle_font

    headers = ["☐", "Tarea", "Zona", "Responsable", "Hora Límite", "Hecha", "Firma"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    done_validation = DataValidation(type="list", formula1='"✓,—"', allow_blank=True)
    ws.add_data_validation(done_validation)
    current_row = 5

    for section_title, zone, tasks in sections:
        ws.merge_cells(f"A{current_row}:G{current_row}")
        cell = ws.cell(row=current_row, column=1, value=f"  {section_title}")
        cell.font = section_font
        zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
        section_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).fill = section_fill
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1

        for task_desc, responsible, time_limit in tasks:
            ws.cell(row=current_row, column=1, value="☐").font = checkbox_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=2, value=task_desc).font = data_font
            ws.cell(row=current_row, column=2).alignment = left_align
            ws.cell(row=current_row, column=2).border = thin_border
            zone_fill = PatternFill(start_color=ZONE_COLORS.get(zone, LIGHT_GRAY),
                                     end_color=ZONE_COLORS.get(zone, LIGHT_GRAY), fill_type="solid")
            ws.cell(row=current_row, column=3, value=zone).font = Font(name="Calibri", size=10, color="666666")
            ws.cell(row=current_row, column=3).fill = zone_fill
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=3).border = thin_border
            ws.cell(row=current_row, column=4, value=responsible).font = data_font
            ws.cell(row=current_row, column=4).fill = input_fill
            ws.cell(row=current_row, column=4).alignment = center_align
            ws.cell(row=current_row, column=4).border = thin_border
            ws.cell(row=current_row, column=5, value=time_limit).font = data_font
            ws.cell(row=current_row, column=5).alignment = center_align
            ws.cell(row=current_row, column=5).border = thin_border
            cell_done = ws.cell(row=current_row, column=6)
            cell_done.fill = input_fill
            cell_done.alignment = center_align
            cell_done.border = thin_border
            done_validation.add(cell_done)
            ws.cell(row=current_row, column=7).fill = input_fill
            ws.cell(row=current_row, column=7).border = thin_border
            current_row += 1
        current_row += 1

    current_row += 1
    ws.merge_cells(f"A{current_row}:C{current_row}")
    ws.cell(row=current_row, column=1, value="Tareas completadas:").font = bold_font
    ws.cell(row=current_row, column=4, value=f'=COUNTIF(F5:F{current_row-2},"✓")').font = bold_font
    ws.cell(row=current_row, column=5, value="de").font = data_font
    total_tasks = sum(len(tasks) for _, _, tasks in sections)
    ws.cell(row=current_row, column=6, value=total_tasks).font = bold_font
    current_row += 2
    ws.cell(row=current_row, column=1, value="Verificado por:").font = bold_font
    ws.merge_cells(f"B{current_row}:D{current_row}")
    ws.cell(row=current_row, column=2).fill = input_fill
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=5, value="Firma:").font = bold_font
    ws.merge_cells(f"F{current_row}:G{current_row}")
    ws.cell(row=current_row, column=6).fill = input_fill
    ws.cell(row=current_row, column=6).border = thin_border
    current_row += 2
    ws.cell(row=current_row, column=1,
            value="— Kit de Tareas Recurrentes · Bar / Cocktails · AI Chef Pro · aichef.pro").font = small_font


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK 1: Apertura y Cierre
# ═══════════════════════════════════════════════════════════════════
def generate_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 — Apertura y Cierre · Bar / Cocktails", [
        "Cómo usar esta plantilla:",
        "▸ Imprime el checklist del turno correspondiente cada día",
        "▸ El responsable de turno marca cada tarea al completarla",
        "▸ Al finalizar el turno, firma y entrega al siguiente responsable",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables — ajusta responsables y horarios",
        "▸ Añade tareas específicas de tu bar en las filas vacías",
        "▸ Borra las tareas que no apliquen a tu negocio",
    ])

    create_task_sheet(wb, "Apertura Barra", "FF8C00", "Apertura Barra Principal", [
        ("Mise en Place de Barra", "Barra", [
            ("Encender iluminación de barra y ambiente", "Bartender", "16:00"),
            ("Encender sistema de música / ambientación sonora", "Bartender", "16:00"),
            ("Verificar limpieza de barra, mostrador y taburetes", "Barback", "16:00"),
            ("Encender cámaras de barra (verificar temperatura 2-4 °C)", "Bartender", "16:00"),
            ("Preparar mise en place de cítricos (rodajas limón, lima, naranja)", "Barback", "16:15"),
            ("Preparar garnish: aceitunas, cerezas, piel de cítricos, hierbas frescas", "Barback", "16:15"),
            ("Llenar cubetas de hielo (cubos, hielo pilé, bloques grandes)", "Barback", "16:00"),
            ("Verificar stock de hielo (mínimo para el turno: 30-50 kg)", "Bartender", "16:00"),
            ("Preparar speed rail con botellas más usadas", "Bartender", "16:15"),
            ("Verificar stock de zumos naturales (preparar si falta)", "Barback", "16:15"),
            ("Preparar jarabes del día (simple, gomme, especiales)", "Bartender", "16:30"),
        ]),
        ("Cristalería y Material", "Barra", [
            ("Pulir cristalería: copas de vino, copas cóctel, vasos old fashioned", "Barback", "16:00"),
            ("Verificar stock de cristalería limpia (mínimo 2x capacidad)", "Barback", "16:15"),
            ("Preparar coctelera, jigger, colador, bar spoon, muddler", "Bartender", "16:15"),
            ("Verificar stock de pajitas, servilletas y posavasos", "Barback", "16:15"),
            ("Preparar tickets de comandas / tablet de pedidos", "Bartender", "16:30"),
        ]),
        ("Máquina de Café", "Café", [
            ("Encender máquina de espresso (precalentar 20 min)", "Bartender", "15:45"),
            ("Purgar grupo (agua caliente 10 segundos por grupo)", "Bartender", "16:00"),
            ("Verificar molienda y calibrar (espresso 25-30 seg, 25-30 ml)", "Bartender", "16:05"),
            ("Verificar stock de leche (entera, avena, soja)", "Barback", "16:00"),
            ("Limpiar boquilla de vapor", "Bartender", "16:05"),
        ]),
        ("Apertura Terraza (si aplica)", "Terraza", [
            ("Sacar mobiliario de terraza (mesas, sillas, sombrillas)", "Barback", "15:30"),
            ("Limpiar mesas y sillas de terraza", "Barback", "15:45"),
            ("Colocar ceniceros, cartas y servilleteros", "Barback", "16:00"),
            ("Verificar iluminación exterior", "Barback", "16:00"),
        ]),
    ])

    create_task_sheet(wb, "Cierre Barra", "8B4513", "Cierre Barra Principal", [
        ("Cierre de Servicio", "Barra", [
            ("Última ronda: avisar a clientes 30 min antes de cierre", "Bartender", "01:30"),
            ("Recoger cristalería de mesas y barra", "Barback", "02:00"),
            ("Vaciar y limpiar cubetas de hielo", "Barback", "02:00"),
            ("Guardar mise en place de cítricos y garnish (film, cámara)", "Bartender", "02:00"),
            ("Guardar jarabes y zumos en cámara (tapar, etiquetar)", "Bartender", "02:00"),
            ("Tapar botellas abiertas de speed rail", "Bartender", "02:00"),
            ("Limpiar coctelera, jigger, colador y utensilios", "Bartender", "02:15"),
        ]),
        ("Limpieza de Barra", "Limpieza", [
            ("Limpiar superficie de barra con desinfectante", "Barback", "02:15"),
            ("Limpiar fregadero de barra (desatascar si necesario)", "Barback", "02:15"),
            ("Limpiar grifo de cerveza (purgar, limpiar boquilla)", "Barback", "02:15"),
            ("Limpiar máquina de espresso (backflush, limpiar grupo)", "Bartender", "02:15"),
            ("Limpiar molinillo de café", "Bartender", "02:15"),
            ("Vaciar y limpiar lavavajillas de barra", "Barback", "02:30"),
            ("Barrer y fregar suelo de barra (detrás de barra)", "Barback", "02:30"),
        ]),
        ("Cierre Administrativo", "Admin", [
            ("Cuadrar caja / cierre de TPV", "Bartender", "02:15"),
            ("Registrar botellas abiertas (inventario rápido)", "Bartender", "02:15"),
            ("Anotar incidencias del turno (rotura, queja, falta stock)", "Bartender", "02:30"),
            ("Apagar música, iluminación decorativa", "Bartender", "02:30"),
            ("Verificar que no queda nadie en el local", "Bartender", "02:30"),
            ("Activar alarma y cerrar con llave", "Bartender", "02:30"),
        ]),
        ("Cierre Terraza", "Terraza", [
            ("Recoger mesas y sillas de terraza (o asegurar)", "Barback", "02:00"),
            ("Recoger ceniceros, cartas y servilleteros", "Barback", "02:00"),
            ("Guardar sombrillas y cojines", "Barback", "02:00"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx"))
    print("✅ 01-apertura-cierre.xlsx")


def generate_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 — Partidas de Barra · Bar / Cocktails", [
        "Tareas organizadas por área de servicio.",
        "▸ Cada pestaña cubre una zona específica del bar",
        "▸ Las tareas están ordenadas cronológicamente",
        "▸ Personaliza con la carta de cócteles de tu bar",
    ])

    create_task_sheet(wb, "Coctelería Clásica", "2196F3", "Partida de Coctelería Clásica", [
        ("Pre-Servicio Coctelería", "Coctelería", [
            ("Verificar stock de spirits base: ginebra, vodka, ron, tequila, whisky, brandy", "Head Bartender", "16:00"),
            ("Verificar stock de vermut (rojo, blanco, dry), amargos y licores", "Head Bartender", "16:00"),
            ("Preparar batch de cócteles de alta rotación (Negroni, Margarita pre-mix)", "Head Bartender", "16:30"),
            ("Preparar infusiones del día (si aplica: gin infusionado, vodka con especias)", "Head Bartender", "16:30"),
            ("Preparar jarabes especiales: miel-jengibre, canela, lavanda, orgeat", "Bartender", "16:30"),
            ("Verificar stock de bitter: Angostura, Peychaud, Orange bitters", "Bartender", "16:15"),
            ("Preparar twist de cítricos frescos (piel de limón, naranja, pomelo)", "Barback", "16:30"),
        ]),
        ("Durante el Servicio", "Coctelería", [
            ("Mantener speed rail organizado y botellas orientadas (etiqueta al frente)", "Bartender", "Continuo"),
            ("Reponer hielo cada hora (o antes si hay pico)", "Barback", "Continuo"),
            ("Reponer garnish y mise en place", "Barback", "Continuo"),
            ("Controlar calidad de cada cóctel antes de servir (prueba visual y aroma)", "Head Bartender", "Continuo"),
            ("Mantener estación limpia entre cócteles (wipe, organizar)", "Bartender", "Continuo"),
        ]),
    ])

    create_task_sheet(wb, "Cerveza y Grifo", "FF9800", "Partida de Cerveza y Grifo", [
        ("Preparación", "Barra", [
            ("Verificar presión de CO₂ de barriles (2.5-3.0 bar)", "Bartender", "16:00"),
            ("Purgar grifos: tirar primer tercio hasta que salga limpio", "Bartender", "16:00"),
            ("Verificar temperatura de barriles en cámara (3-5 °C)", "Bartender", "16:00"),
            ("Verificar stock de barriles de reserva por marca", "Bartender", "16:15"),
            ("Limpiar boquillas de grifo con agua caliente", "Barback", "16:00"),
            ("Preparar cristalería de cerveza (congelar copas si es estilo del bar)", "Barback", "16:00"),
        ]),
        ("Control de Stock de Botella", "Bodega", [
            ("Verificar stock de cerveza en botella y lata en cámara", "Barback", "16:00"),
            ("Reponer cámara de barra con stock de bodega", "Barback", "16:15"),
            ("Verificar stock de refrescos: tónica, soda, cola, ginger ale, ginger beer", "Barback", "16:15"),
            ("Verificar stock de agua mineral y con gas", "Barback", "16:15"),
        ]),
    ])

    create_task_sheet(wb, "Vinos y Bodega", "9C27B0", "Partida de Vinos por Copa y Bodega", [
        ("Vinos por Copa", "Bodega", [
            ("Verificar vinos abiertos por copa (blancos y rosados en cámara a 8-10 °C)", "Bartender", "16:00"),
            ("Verificar vinos tintos abiertos (temperatura ambiente o 16-18 °C)", "Bartender", "16:00"),
            ("Comprobar que ningún vino abierto tiene más de 3 días (oler, probar)", "Bartender", "16:00"),
            ("Abrir botellas nuevas si quedan menos de 2 copas de algún vino", "Bartender", "16:15"),
            ("Actualizar pizarra de vinos por copa (si cambió alguno)", "Bartender", "16:30"),
        ]),
        ("Bodega y Reservas", "Bodega", [
            ("Verificar temperatura de bodega / vinoteca (12-16 °C)", "Bartender", "16:00"),
            ("Revisar stock de vinos de carta (marcar los agotados)", "Bartender", "Semanal"),
            ("Organizar botellas por zona/denominación", "Barback", "Semanal"),
            ("Revisar estado de vinos especiales / añadas", "Head Bartender", "Mensual"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "02-partidas-cocina.xlsx"))
    print("✅ 02-partidas-cocina.xlsx")


def generate_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 — Tareas del Manager · Bar / Cocktails", [
        "Checklists para el/la responsable del bar.",
        "▸ Diario: lo que hay que revisar cada día",
        "▸ Semanal: tareas organizadas por día",
        "▸ Mensual: revisiones periódicas",
    ])

    create_task_sheet(wb, "Diario Manager", "FFD700", "Checklist Diario del Manager / Responsable", [
        ("Antes de Abrir", "Admin", [
            ("Revisar reservas del día (grupos, eventos privados, VIPs)", "Manager", "15:00"),
            ("Verificar asistencia del equipo y cubrir bajas", "Manager", "15:00"),
            ("Revisar stock crítico (spirits, cerveza, hielo, cítricos)", "Manager", "15:15"),
            ("Confirmar entregas de proveedores del día", "Manager", "15:15"),
            ("Briefing con equipo: cóctel del día, eventos, notas", "Manager", "15:45"),
        ]),
        ("Durante el Servicio", "Admin", [
            ("Supervisar calidad de servicio y cócteles", "Manager", "Continuo"),
            ("Control de ambiente: música, iluminación, temperatura", "Manager", "Continuo"),
            ("Gestionar quejas o incidencias con clientes", "Manager", "Continuo"),
            ("Supervisar cumplimiento de normativa (no servir a menores, aforo)", "Manager", "Continuo"),
            ("Evaluar tiempos de servicio (objetivo: <3 min cóctel, <1 min cerveza)", "Manager", "Continuo"),
        ]),
        ("Cierre de Día", "Admin", [
            ("Revisar caja y cuadrar con TPV", "Manager", "Cierre"),
            ("Revisar inventario rápido de botellas abiertas", "Manager", "Cierre"),
            ("Anotar mermas (rotura cristalería, derrames, producto caducado)", "Manager", "Cierre"),
            ("Planificar necesidades del día siguiente", "Manager", "Cierre"),
            ("Verificar cierre correcto del local", "Manager", "Cierre"),
        ]),
    ])

    create_task_sheet(wb, "Semanal Manager", "FF8C00", "Checklist Semanal del Manager", [
        ("Lunes — Inventario", "Admin", [
            ("Inventario completo de spirits (botellas llenas + nivel de abiertas)", "Manager", "15:00"),
            ("Inventario de cerveza (barriles + botellas + latas)", "Manager", "15:30"),
            ("Inventario de vinos (bodega + cámara)", "Manager", "16:00"),
            ("Inventario de refrescos, zumos, jarabes y mixers", "Manager", "16:15"),
            ("Hacer pedidos semanales a distribuidores", "Manager", "16:30"),
        ]),
        ("Martes — Equipo", "Admin", [
            ("Revisar cuadrante de turnos de la semana", "Manager", "15:00"),
            ("Reunión con equipo (feedback, formación, objetivos)", "Manager", "15:30"),
            ("Revisar formación de personal nuevo (mentoring)", "Manager", "16:00"),
        ]),
        ("Miércoles — Calidad", "Admin", [
            ("Auditar 3 cócteles al azar (receta, presentación, tiempo)", "Manager", "21:00"),
            ("Revisar registros de temperatura de cámaras", "Manager", "16:00"),
            ("Comprobar FIFO en cámaras y bodega", "Manager", "16:00"),
            ("Verificar limpieza profunda de barra (detrás, debajo)", "Manager", "16:30"),
        ]),
        ("Jueves — Carta y Creatividad", "Admin", [
            ("Revisar ventas por cóctel (top 5 y bottom 5)", "Manager", "15:00"),
            ("Desarrollar o rotar cóctel de la semana", "Head Bartender", "15:30"),
            ("Actualizar carta si hay cambios", "Manager", "16:00"),
            ("Preparar contenido RRSS (foto de cóctel estrella)", "Manager", "16:30"),
        ]),
        ("Viernes — Pre Fin de Semana", "Admin", [
            ("Verificar stock para fin de semana (30-50% más que entre semana)", "Manager", "15:00"),
            ("Confirmar turnos reforzados de fin de semana", "Manager", "15:15"),
            ("Verificar stock de hielo extra", "Manager", "15:15"),
            ("Preparar batches de cócteles de alta rotación", "Head Bartender", "16:00"),
        ]),
    ])

    create_task_sheet(wb, "Mensual Manager", "8B4513", "Checklist Mensual del Manager", [
        ("Financiero", "Admin", [
            ("Revisar pour cost (objetivo: 18-22% spirits, 25-30% cerveza)", "Manager", "1ª semana"),
            ("Analizar mermas del mes (objetivo: <3%)", "Manager", "1ª semana"),
            ("Comparar ventas mes actual vs. anterior y mismo mes año anterior", "Manager", "1ª semana"),
            ("Revisar rentabilidad por categoría (cócteles, cerveza, vino, café)", "Manager", "1ª semana"),
        ]),
        ("Mantenimiento", "Admin", [
            ("Limpieza profunda de líneas de grifo (con producto específico)", "Mantenimiento", "2ª semana"),
            ("Mantenimiento de máquina de espresso (descalcificar, cambiar juntas)", "Mantenimiento", "2ª semana"),
            ("Revisar estado de cámaras frigoríficas (limpieza, temperatura)", "Mantenimiento", "2ª semana"),
            ("Verificar estado de cristalería (reponer piezas rotas/rayadas)", "Manager", "2ª semana"),
            ("Calibrar termómetros y sondas", "Manager", "2ª semana"),
        ]),
        ("Evaluación", "Admin", [
            ("Evaluación informal del equipo (skills, actitud, puntualidad)", "Manager", "3ª semana"),
            ("Visitar 2-3 bares de la competencia (tendencias, precios, servicio)", "Manager", "3ª semana"),
            ("Proponer nuevos cócteles para la carta", "Head Bartender", "3ª semana"),
            ("Revisar reviews online (Google, TripAdvisor, redes)", "Manager", "4ª semana"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "03-tareas-manager.xlsx"))
    print("✅ 03-tareas-manager.xlsx")


def generate_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 — Tareas por Perfil · Bar / Cocktails", [
        "Checklists específicos para cada perfil profesional.",
        "▸ Cada pestaña = un puesto de trabajo",
        "▸ Perfecto para onboarding de personal nuevo",
    ])

    create_task_sheet(wb, "Head Bartender", "FFD700", "Tareas Diarias — Head Bartender", [
        ("Pre-Servicio", "Coctelería", [
            ("Verificar mise en place completa de coctelería", "Head Bartender", "16:00"),
            ("Preparar cócteles especiales del día / de la carta secreta", "Head Bartender", "16:30"),
            ("Supervisar calibración de café (espresso test)", "Head Bartender", "16:00"),
            ("Briefing con equipo: cóctel del día, VIPs, eventos", "Head Bartender", "16:45"),
            ("Verificar que la carta de cócteles está actualizada", "Head Bartender", "16:30"),
        ]),
        ("Servicio", "Coctelería", [
            ("Preparar cócteles premium y de alta complejidad", "Head Bartender", "Continuo"),
            ("Supervisar calidad de cócteles preparados por el equipo", "Head Bartender", "Continuo"),
            ("Atender a clientes VIP y regulares en barra", "Head Bartender", "Continuo"),
            ("Recomendar cócteles según preferencias del cliente", "Head Bartender", "Continuo"),
            ("Controlar tiempos de preparación", "Head Bartender", "Continuo"),
        ]),
        ("Cierre y Desarrollo", "Admin", [
            ("Supervisar cierre correcto de barra", "Head Bartender", "Cierre"),
            ("Desarrollar nuevas recetas y técnicas", "Head Bartender", "Continuo"),
            ("Formar al equipo en nuevas técnicas", "Head Bartender", "Semanal"),
            ("Calcular coste de nuevos cócteles (escandallo)", "Head Bartender", "Semanal"),
        ]),
    ])

    create_task_sheet(wb, "Bartender", "FF8C00", "Tareas Diarias — Bartender", [
        ("Pre-Servicio", "Barra", [
            ("Preparar mise en place de su estación", "Bartender", "16:00"),
            ("Cortar cítricos y preparar garnish", "Bartender", "16:15"),
            ("Preparar jarabes del día", "Bartender", "16:30"),
            ("Pulir cristalería de su zona", "Bartender", "16:15"),
            ("Verificar stock de su speed rail", "Bartender", "16:15"),
        ]),
        ("Servicio", "Barra", [
            ("Preparar cócteles según carta y pedidos especiales", "Bartender", "Continuo"),
            ("Servir cerveza de grifo y botella", "Bartender", "Continuo"),
            ("Servir vinos por copa", "Bartender", "Continuo"),
            ("Preparar cafés e infusiones", "Bartender", "Continuo"),
            ("Cobrar en barra (si el sistema lo permite)", "Bartender", "Continuo"),
            ("Mantener estación limpia y ordenada", "Bartender", "Continuo"),
        ]),
        ("Cierre", "Barra", [
            ("Limpiar estación completa", "Bartender", "Cierre"),
            ("Guardar mise en place en cámara", "Bartender", "Cierre"),
            ("Lavar y guardar utensilios personales", "Bartender", "Cierre"),
        ]),
    ])

    create_task_sheet(wb, "Barback", "4CAF50", "Tareas Diarias — Barback / Auxiliar de Barra", [
        ("Soporte Continuo", "Barra", [
            ("Reponer hielo constantemente (antes de que pidan)", "Barback", "Continuo"),
            ("Reponer cristalería limpia en estaciones", "Barback", "Continuo"),
            ("Recoger cristalería sucia de mesas y barra", "Barback", "Continuo"),
            ("Lavar cristalería en lavavajillas de barra", "Barback", "Continuo"),
            ("Reponer servilletas, pajitas y posavasos", "Barback", "Continuo"),
            ("Reponer botellas de speed rail desde bodega", "Barback", "Continuo"),
            ("Reponer refrescos en cámara de barra", "Barback", "Continuo"),
            ("Mantener suelo limpio (secar derrames inmediatamente)", "Barback", "Continuo"),
        ]),
        ("Limpieza", "Limpieza", [
            ("Limpiar mesas y recoger vasos de zona de clientes", "Barback", "Continuo"),
            ("Limpiar baños cada 2 horas (reponer jabón, papel)", "Barback", "Cada 2h"),
            ("Vaciar papeleras y ceniceros de terraza", "Barback", "Continuo"),
            ("Barrer y fregar al cierre", "Barback", "Cierre"),
            ("Sacar basura y reciclar vidrio", "Barback", "Cierre"),
        ]),
    ])

    create_task_sheet(wb, "Camarero Barra", "2196F3", "Tareas Diarias — Camarero de Sala / Terraza", [
        ("Servicio", "Terraza", [
            ("Atender mesas de sala y/o terraza", "Camarero", "Continuo"),
            ("Tomar comandas con conocimiento de la carta de cócteles", "Camarero", "Continuo"),
            ("Servir bebidas desde barra a mesa", "Camarero", "Continuo"),
            ("Recomendar cócteles y maridajes", "Camarero", "Continuo"),
            ("Cobrar en mesa (datáfono portátil)", "Camarero", "Continuo"),
            ("Recoger cristalería y mantener mesas limpias", "Camarero", "Continuo"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "04-tareas-perfiles.xlsx"))
    print("✅ 04-tareas-perfiles.xlsx")


def generate_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 — Semanales y Mensuales · Bar / Cocktails", [
        "Tareas de frecuencia semanal o mensual.",
        "▸ Limpieza profunda por zona",
        "▸ Mantenimiento de equipos",
        "▸ Inventario y control de stock",
    ])

    create_task_sheet(wb, "Limpieza Semanal", "E91E63", "Limpieza Profunda Semanal", [
        ("Barra y Coctelería", "Limpieza", [
            ("Limpiar interior de todas las cámaras de barra", "Barback", "Lunes"),
            ("Limpiar grifos de cerveza con producto específico", "Barback", "Lunes"),
            ("Limpiar interior de lavavajillas de barra", "Barback", "Lunes"),
            ("Limpiar estanterías de botellas (polvo, derrames)", "Barback", "Martes"),
            ("Limpiar speed rail y pouring spouts", "Barback", "Martes"),
            ("Limpiar espejo y cristales de barra", "Barback", "Miércoles"),
            ("Fregar suelo detrás de barra con desengrasante", "Barback", "Jueves"),
            ("Limpiar campana extractora (si hay cocina de barra)", "Barback", "Viernes"),
        ]),
        ("Sala y Terraza", "Limpieza", [
            ("Limpiar tapicería de taburetes y sillas", "Barback", "Miércoles"),
            ("Limpiar mesas a fondo (patas, base)", "Barback", "Miércoles"),
            ("Limpiar cristales y espejos de sala", "Barback", "Jueves"),
            ("Limpiar mobiliario de terraza a fondo", "Barback", "Viernes"),
        ]),
        ("Baños", "Limpieza", [
            ("Limpieza profunda de baños (suelos, sanitarios, espejos)", "Barback", "Lunes"),
            ("Verificar funcionamiento de cisternas y grifos", "Manager", "Lunes"),
            ("Reponer dispensadores de jabón y papel", "Barback", "Lunes"),
        ]),
    ])

    create_task_sheet(wb, "Mantenimiento Mensual", "FF9800", "Mantenimiento Mensual de Equipos", [
        ("Grifo y Cerveza", "Barra", [
            ("Limpieza completa de líneas de grifo (producto profesional)", "Mantenimiento", "1ª semana"),
            ("Verificar presión y reguladores de CO₂", "Mantenimiento", "1ª semana"),
            ("Revisar juntas y acoplamientos de barriles", "Mantenimiento", "1ª semana"),
        ]),
        ("Café", "Café", [
            ("Descalcificar máquina de espresso", "Mantenimiento", "2ª semana"),
            ("Cambiar juntas de grupo si es necesario", "Mantenimiento", "2ª semana"),
            ("Calibrar presión de la máquina (9 bar)", "Mantenimiento", "2ª semana"),
            ("Limpiar molinillo (vaciar, aspirar, calibrar)", "Bartender", "2ª semana"),
        ]),
        ("Frío y General", "Barra", [
            ("Limpiar condensadores de cámaras", "Mantenimiento", "3ª semana"),
            ("Verificar temperaturas con sonda independiente", "Manager", "3ª semana"),
            ("Revisar juntas de puertas de cámaras", "Mantenimiento", "3ª semana"),
            ("Inventario de cristalería (reponer rotas/rayadas)", "Manager", "3ª semana"),
        ]),
    ])

    create_task_sheet(wb, "Inventario", "4CAF50", "Inventario Semanal y Control de Stock", [
        ("Spirits y Licores", "Bodega", [
            ("Contar botellas llenas + medir nivel de abiertas: ginebra", "Bartender", "Lunes"),
            ("Contar botellas: vodka, ron blanco, ron oscuro", "Bartender", "Lunes"),
            ("Contar botellas: tequila, mezcal, whisky/bourbon", "Bartender", "Lunes"),
            ("Contar botellas: brandy, cognac, licores (Cointreau, Campari, Aperol...)", "Bartender", "Lunes"),
            ("Contar botellas: vermut (rojo, blanco, dry), amargos", "Bartender", "Lunes"),
        ]),
        ("Cerveza, Vino y Refrescos", "Bodega", [
            ("Contar barriles: llenos + porcentaje de los conectados", "Barback", "Lunes"),
            ("Contar cerveza botella/lata por marca", "Barback", "Lunes"),
            ("Contar botellas de vino en bodega y cámara", "Bartender", "Lunes"),
            ("Contar stock de tónica, soda, cola, ginger ale, ginger beer, zumos", "Barback", "Lunes"),
            ("Contar stock de agua mineral y con gas", "Barback", "Lunes"),
        ]),
        ("Consumibles", "Barra", [
            ("Contar stock de hielo (o verificar producción de máquina)", "Barback", "Lunes"),
            ("Contar stock de cítricos (limones, limas, naranjas)", "Barback", "Lunes"),
            ("Contar stock de servilletas, pajitas, posavasos", "Barback", "Lunes"),
            ("Contar stock de café en grano y leches", "Bartender", "Lunes"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx"))
    print("✅ 05-tareas-semanales-mensuales.xlsx")


def generate_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 — Eventos y Festivos · Bar / Cocktails", [
        "Checklists para noches especiales y eventos.",
        "▸ Los bares facturan su máximo en noches de evento y temporadas festivas",
        "▸ Personaliza con los eventos de tu zona",
    ])

    create_task_sheet(wb, "Noche Cócteles", "9C27B0", "Noche de Cócteles Especiales / Tasting", [
        ("Preparación (48h antes)", "Admin", [
            ("Diseñar menú de cócteles especiales (4-6 cócteles)", "Head Bartender", "48h antes"),
            ("Calcular ingredientes necesarios y hacer pedidos extra", "Manager", "48h antes"),
            ("Preparar infusiones, cordials o jarabes especiales", "Head Bartender", "24h antes"),
            ("Publicar evento en RRSS con fotos de cócteles", "Manager", "48h antes"),
        ]),
        ("Día del Evento", "Coctelería", [
            ("Preparar batches de cócteles especiales", "Head Bartender", "16:00"),
            ("Preparar garnish especial (deshidratados, flores comestibles, espuma)", "Bartender", "16:30"),
            ("Imprimir menú especial / colocar en mesas", "Manager", "17:00"),
            ("Reforzar equipo (bartender extra, barback extra)", "Manager", "Apertura"),
            ("Stock extra de hielo (50% más que un día normal)", "Barback", "16:00"),
        ]),
    ])

    create_task_sheet(wb, "Nochevieja", "C62828", "Nochevieja / Fin de Año", [
        ("Planificación", "Admin", [
            ("Definir concepto de la noche (fiesta, cena + barra libre, precio entrada)", "Manager", "3 sem antes"),
            ("Calcular stock necesario (3-4x un viernes normal)", "Manager", "2 sem antes"),
            ("Contratar personal extra", "Manager", "2 sem antes"),
            ("Preparar decoración especial", "Manager", "1 sem antes"),
            ("Preparar uvas y cava/champagne para las 12", "Manager", "Día antes"),
        ]),
        ("Día de Nochevieja", "Barra", [
            ("Stock masivo de hielo", "Barback", "Mañana"),
            ("Preparar batches de cócteles de alta rotación", "Head Bartender", "Tarde"),
            ("Preparar botellas de champagne/cava en hielo", "Barback", "22:00"),
            ("Servir cava a medianoche", "Todo equipo", "00:00"),
            ("Gestionar aforo y seguridad", "Manager", "Continuo"),
        ]),
    ])

    create_task_sheet(wb, "After Work", "FF6F00", "Sesión After-Work / Happy Hour", [
        ("Preparación", "Admin", [
            ("Definir horario de happy hour (ej. 18:00-20:00)", "Manager", "Fijo"),
            ("Definir carta reducida de cócteles con descuento", "Head Bartender", "Fijo"),
            ("Publicar en RRSS y colocar pizarra en la entrada", "Manager", "Diario"),
        ]),
        ("Servicio After-Work", "Barra", [
            ("Preparar batches de cócteles de happy hour", "Bartender", "17:30"),
            ("Preparar snacks / tapas de cortesía (si aplica)", "Barback", "17:30"),
            ("Servicio rápido: objetivo <2 min por cóctel", "Bartender", "Continuo"),
            ("Mantener música y ambiente apropiados para after-work", "Manager", "18:00"),
        ]),
    ])

    create_task_sheet(wb, "Maridaje Catas", "4CAF50", "Maridaje / Cata Guiada", [
        ("Preparación", "Admin", [
            ("Seleccionar 4-6 piezas para cata (vinos, spirits, cócteles)", "Head Bartender", "1 sem antes"),
            ("Preparar fichas de cata para clientes", "Manager", "3 días antes"),
            ("Confirmar asistentes y preparar sala", "Manager", "1 día antes"),
            ("Preparar maridaje food (quesos, chocolates, tapas)", "Manager", "Día"),
        ]),
        ("Evento", "Coctelería", [
            ("Presentar cada pieza: historia, notas de cata, técnica", "Head Bartender", "Evento"),
            ("Servir en copa/vaso correcto", "Bartender", "Evento"),
            ("Gestionar ritmo de la cata (15-20 min por pieza)", "Head Bartender", "Evento"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "06-eventos-festivos.xlsx"))
    print("✅ 06-eventos-festivos.xlsx")


def generate_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 — Plantilla Personalizable · Bar / Cocktails", [
        "Plantillas en blanco para crear tus propias listas de tareas.",
        "",
        "Cómo usar:",
        "▸ Copia la pestaña que más se ajuste",
        "▸ Rellena con las tareas específicas de tu bar",
        "▸ Imprime en A4 y distribuye al equipo",
    ])

    create_task_sheet(wb, "Por Franja Horaria", "FFD700", "Plantilla en Blanco — Por Franja Horaria", [
        ("Apertura (15:00 - 18:00)", "General", [("", "", "")] * 5),
        ("Servicio Tarde (18:00 - 22:00)", "General", [("", "", "")] * 5),
        ("Servicio Noche (22:00 - 02:00)", "General", [("", "", "")] * 5),
        ("Cierre (02:00 - 03:00)", "General", [("", "", "")] * 4),
    ])

    create_task_sheet(wb, "Por Zona", "FF8C00", "Plantilla en Blanco — Por Zona", [
        ("Barra Principal", "Barra", [("", "", "")] * 5),
        ("Coctelería", "Coctelería", [("", "", "")] * 5),
        ("Bodega / Almacén", "Bodega", [("", "", "")] * 4),
        ("Terraza", "Terraza", [("", "", "")] * 4),
    ])

    create_task_sheet(wb, "Por Perfil", "9C27B0", "Plantilla en Blanco — Por Perfil", [
        ("Head Bartender", "General", [("", "", "")] * 4),
        ("Bartender", "General", [("", "", "")] * 4),
        ("Barback", "General", [("", "", "")] * 4),
        ("Camarero", "General", [("", "", "")] * 4),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "07-plantilla-personalizable.xlsx"))
    print("✅ 07-plantilla-personalizable.xlsx")


def generate_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 01 — Briefing de Servicio · Bar", [
        "Plantilla para la reunión diaria del equipo antes de abrir.",
        "▸ 5 minutos que mejoran la comunicación",
        "▸ Imprime y pega en barra cada día",
    ])

    create_task_sheet(wb, "Briefing Diario", "FFD700", "Briefing Pre-Servicio — Bar / Cocktails", [
        ("Servicio del Día", "Barra", [
            ("Cóctel del día / cóctel especial: nombre, ingredientes, precio", "Head Bartender", "Pre-apertura"),
            ("Reservas y eventos privados del día", "Manager", "Pre-apertura"),
            ("Clientes VIP o regulares que se esperan", "Manager", "Pre-apertura"),
            ("Promociones activas (happy hour, 2x1, etc.)", "Manager", "Pre-apertura"),
        ]),
        ("Alertas y Notas", "General", [
            ("Productos agotados o en falta", "Manager", "Pre-apertura"),
            ("Cambios en la carta o precios", "Manager", "Pre-apertura"),
            ("Incidencias del turno anterior", "Manager", "Pre-apertura"),
            ("Cumpleaños o celebraciones de clientes (si se sabe)", "Manager", "Pre-apertura"),
        ]),
        ("Equipo del Turno", "General", [
            ("Quién trabaja hoy y en qué puesto (barra, sala, barback)", "Manager", "Pre-apertura"),
            ("Personal nuevo / en formación (asignar mentor)", "Manager", "Pre-apertura"),
            ("Horarios especiales de hoy", "Manager", "Pre-apertura"),
        ]),
    ])

    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))
    print("✅ BONUS-01-briefing-servicio.xlsx")


def generate_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 02 — Calendario Anual · Bar / Cocktails", [
        "Fechas clave para bares con tareas asociadas.",
        "▸ Cada fecha incluye qué preparar y con cuánta antelación",
        "▸ Añade las fechas locales de tu zona",
    ])

    ws = wb.create_sheet(title="Calendario Anual")
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Calendario Anual de Fechas Clave — Bar / Cocktails"
    ws["A1"].font = title_font

    headers = ["#", "Fecha", "Evento", "Preparación Especial", "Antelación", "Notas"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    events = [
        ("1", "14 Febrero", "San Valentín", "Cócteles para parejas, carta especial, decoración", "2 semanas", "Reforzar servicio de mesa"),
        ("2", "17 Marzo", "St. Patrick's Day", "Stock extra de cerveza, whisky irlandés, decoración verde", "2 semanas", "Evento muy potente para bares"),
        ("3", "Mar-Abr", "Semana Santa", "Ajustar horarios, cócteles de temporada", "1 semana", "Variable según zona"),
        ("4", "5 Mayo", "Cinco de Mayo", "Margaritas, mezcal, tequila, decoración mexicana", "2 semanas", "Tendencia en crecimiento"),
        ("5", "21 Junio", "Solsticio / Inicio Verano", "Cócteles refrescantes, terraza a tope", "1 semana", "Abrir temporada de terraza"),
        ("6", "Jun-Sep", "Temporada Verano", "Mojitos, gin tonics, spritz, cócteles con fruta", "4 semanas", "Stock extra de hielo y cítricos"),
        ("7", "31 Octubre", "Halloween", "Cócteles temáticos, decoración, hielo seco", "2 semanas", "Noche fuerte de facturación"),
        ("8", "Noviembre", "Beaujolais Nouveau", "Vino nuevo, maridaje", "1 semana", "Para bares de vino"),
        ("9", "Diciembre", "Navidad/Empresas", "Cenas de empresa, barra libre, carta especial", "4 semanas", "Pico máximo del año para bares"),
        ("10", "31 Diciembre", "Nochevieja", "Champagne, cava, fiesta, entrada con consumición", "3 semanas", "La noche MÁS importante"),
        ("11", "Variable", "Fiestas Locales", "Adaptar carta y horarios a fiestas patronales", "3 semanas", "AÑADE TUS FECHAS LOCALES"),
        ("12", "Variable", "Eventos Deportivos", "Stock extra cerveza, pantallas, horarios especiales", "1 semana", "Champions, Eurocopa, Mundial"),
        ("13", "Variable", "World Cocktail Day (13 May)", "Cóctel gratis, taller, promo RRSS", "2 semanas", "Marketing y branding"),
        ("14", "Variable", "Negroni Week (Sep)", "Carta de Negronis, donación benéfica por cóctel", "2 semanas", "Evento global de coctelería"),
        ("15", "Variable", "After-Work Semanal", "Happy hour recurrente, fidelización", "Fijo", "Programa de fidelización"),
        ("16", "Variable", "Catas y Maridajes", "Cata guiada mensual (vinos, spirits, cócteles)", "3 semanas", "Diferenciación y premium"),
        ("17", "Variable", "DJ / Música en Vivo", "Sonido, decoración, stock extra, aforo", "2 semanas", "Control de ruido vecinal"),
    ]

    for i, (num, fecha, evento, produccion, antelacion, notas) in enumerate(events, start=4):
        ws.cell(row=i, column=1, value=num).font = data_font
        ws.cell(row=i, column=1).alignment = center_align
        ws.cell(row=i, column=1).border = thin_border
        ws.cell(row=i, column=2, value=fecha).font = bold_font
        ws.cell(row=i, column=2).alignment = center_align
        ws.cell(row=i, column=2).border = thin_border
        ws.cell(row=i, column=3, value=evento).font = bold_font
        ws.cell(row=i, column=3).border = thin_border
        ws.cell(row=i, column=4, value=produccion).font = data_font
        ws.cell(row=i, column=4).alignment = left_align
        ws.cell(row=i, column=4).border = thin_border
        ws.cell(row=i, column=5, value=antelacion).font = data_font
        ws.cell(row=i, column=5).alignment = center_align
        ws.cell(row=i, column=5).border = thin_border
        ws.cell(row=i, column=6, value=notas).font = Font(name="Calibri", size=10, color="666666")
        ws.cell(row=i, column=6).alignment = left_align
        ws.cell(row=i, column=6).border = thin_border
        if "AÑADE" in notas:
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = input_fill

    footer_row = len(events) + 5
    ws.cell(row=footer_row, column=1,
            value="— Kit de Tareas Recurrentes · Bar / Cocktails · AI Chef Pro · aichef.pro").font = small_font

    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx"))
    print("✅ BONUS-02-calendario-anual-tareas.xlsx")


if __name__ == "__main__":
    print("\n🍸 Generando Kit de Tareas Recurrentes — Bar / Cocktails\n")
    generate_01()
    generate_02()
    generate_03()
    generate_04()
    generate_05()
    generate_06()
    generate_07()
    generate_bonus_01()
    generate_bonus_02()
    print(f"\n✅ 9 archivos generados en {OUTPUT_DIR}\n")
