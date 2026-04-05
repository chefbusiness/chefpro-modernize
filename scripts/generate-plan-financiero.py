#!/usr/bin/env python3
"""
Generate Kit Plan Financiero para Restaurantes (9 archivos).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-plan-financiero"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Semaphore
VERDE = "C8E6C9"
AMARILLO = "FFF9C4"
ROJO = "FFCDD2"

# Input highlight
INPUT_COLOR = "E8F5E9"

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
formula_font = Font(name="Calibri", size=11, color="1565C0", bold=True)
pct_font = Font(name="Calibri", size=11, color="1565C0")

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
input_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
verde_fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
amarillo_fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
rojo_fill = PatternFill(start_color=ROJO, end_color=ROJO, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

BRAND_LINE = "AI Chef Pro · aichef.pro — Kit Plan Financiero para Restaurantes"
EUR_FMT = '#,##0.00 €'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'

MONTHS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 85
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    ws["B4"].value = BRAND_LINE
    ws["B4"].font = subtitle_font
    for i, line in enumerate(lines, start=6):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def style_header_row(ws, row, headers, col_start=1):
    for col_idx, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def add_brand_footer(ws, row, merge_end="H"):
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


def set_col_widths(ws, widths):
    if isinstance(widths, dict):
        widths = widths.items()
    for letter, w in widths:
        ws.column_dimensions[letter].width = w


def style_data_cell(ws, row, col, value=None, fmt=None, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or data_font
    cell.border = thin_border
    cell.alignment = align or right_align
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


def alt_fill(row):
    return light_fill if row % 2 == 0 else None


# ═══════════════════════════════════════════════════════════
# 01 — PLAN FINANCIERO PREVISIONAL (P&L 3 AÑOS)
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Plan Financiero Previsional — P&L 3 Años", [
        "Cómo usar esta plantilla:",
        "▸ Completa los datos mensuales de ingresos y gastos en cada pestaña de año.",
        "▸ Las celdas en verde son campos editables (inputs).",
        "▸ EBITDA y márgenes se calculan automáticamente.",
        "▸ La pestaña Resumen consolida los 3 años para presentación a inversores o bancos.",
        "",
        "Líneas de ingreso:",
        "▸ Comedor, Barra, Delivery, Eventos/Catering.",
        "",
        "Líneas de gasto:",
        "▸ Food Cost, Personal, Alquiler, Suministros, Marketing, Administración.",
    ])

    revenue_lines = ["Comedor", "Barra", "Delivery", "Eventos / Catering"]
    cost_lines = ["Food Cost", "Personal (Labor)", "Alquiler", "Suministros", "Marketing", "Administración / Otros"]

    for year_num in range(1, 4):
        ws = wb.create_sheet(f"Año {year_num}")
        ws.sheet_properties.tabColor = "1565C0"
        set_col_widths(ws, [("A", 28)] + [(get_column_letter(i), 14) for i in range(2, 15)])

        ws.merge_cells("A1:N1")
        ws["A1"].value = f"P&L Previsional — Año {year_num}"
        ws["A1"].font = title_font
        ws.merge_cells("A2:N2")
        ws["A2"].value = BRAND_LINE
        ws["A2"].font = subtitle_font

        headers = ["Concepto"] + MONTHS + ["TOTAL AÑO"]
        style_header_row(ws, 4, headers)

        # INGRESOS section
        r = 5
        ws.cell(row=r, column=1, value="INGRESOS").font = section_font
        r += 1
        rev_start = r
        for line in revenue_lines:
            ws.cell(row=r, column=1, value=line).font = data_font
            ws.cell(row=r, column=1).border = thin_border
            for c in range(2, 14):
                style_data_cell(ws, r, c, 0, EUR_FMT, fill=input_fill)
            # Total formula
            style_data_cell(ws, r, 14, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
            r += 1
        rev_end = r - 1

        # Total Ingresos
        ws.cell(row=r, column=1, value="TOTAL INGRESOS").font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=c).value = f"=SUM({col_l}{rev_start}:{col_l}{rev_end})"
        total_rev_row = r
        r += 2

        # GASTOS section
        ws.cell(row=r, column=1, value="GASTOS").font = section_font
        r += 1
        cost_start = r
        for line in cost_lines:
            ws.cell(row=r, column=1, value=line).font = data_font
            ws.cell(row=r, column=1).border = thin_border
            for c in range(2, 14):
                style_data_cell(ws, r, c, 0, EUR_FMT, fill=input_fill)
            style_data_cell(ws, r, 14, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
            r += 1
        cost_end = r - 1

        # Total Gastos
        ws.cell(row=r, column=1, value="TOTAL GASTOS").font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=c).value = f"=SUM({col_l}{cost_start}:{col_l}{cost_end})"
        total_cost_row = r
        r += 2

        # EBITDA
        ws.cell(row=r, column=1, value="EBITDA").font = Font(name="Calibri", size=13, bold=True, color=GOLD)
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=EUR_FMT, font=Font(name="Calibri", size=12, bold=True, color=GOLD))
            ws.cell(row=r, column=c).value = f"={col_l}{total_rev_row}-{col_l}{total_cost_row}"
        ebitda_row = r
        r += 1

        # Margen EBITDA %
        ws.cell(row=r, column=1, value="Margen EBITDA %").font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=PCT_FMT, font=pct_font)
            ws.cell(row=r, column=c).value = f'=IF({col_l}{total_rev_row}=0,0,{col_l}{ebitda_row}/{col_l}{total_rev_row})'
        r += 2
        add_brand_footer(ws, r, "N")

    # Resumen tab
    ws = wb.create_sheet("Resumen")
    ws.sheet_properties.tabColor = GOLD
    set_col_widths(ws, [("A", 28), ("B", 18), ("C", 18), ("D", 18), ("E", 18)])

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Resumen Ejecutivo — P&L 3 Años"
    ws["A1"].font = title_font
    ws.merge_cells("A2:E2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Concepto", "Año 1", "Año 2", "Año 3", "Crecimiento Y1→Y3"]
    style_header_row(ws, 4, headers)

    summary_items = ["Total Ingresos", "Total Gastos", "EBITDA", "Margen EBITDA %"]
    for i, item in enumerate(summary_items, 5):
        ws.cell(row=i, column=1, value=item).font = bold_font
        ws.cell(row=i, column=1).border = thin_border
        for c in range(2, 6):
            style_data_cell(ws, i, c, 0, EUR_FMT if "%" not in item else PCT_FMT)
    ws.cell(row=5, column=5).value = "=IF(B5=0,0,(D5-B5)/B5)"
    ws.cell(row=5, column=5).number_format = PCT_FMT

    add_brand_footer(ws, 12, "E")

    wb.save(os.path.join(OUTPUT_DIR, "01-plan-financiero-previsional.xlsx"))
    print("  ✓ 01-plan-financiero-previsional.xlsx")


# ═══════════════════════════════════════════════════════════
# 01b — PLAN FINANCIERO PREVISIONAL (P&L 5 AÑOS)
# ═══════════════════════════════════════════════════════════
def gen_01b():
    wb = Workbook()
    add_instructions_sheet(wb, "01b · Plan Financiero Previsional — P&L 5 Años", [
        "Cómo usar esta plantilla:",
        "▸ Idéntica estructura que el plan a 3 años, pero con proyección a 5 años.",
        "▸ Ideal para presentaciones a bancos, inversores o franquicias que requieren 5 años.",
        "▸ Año 1 con desglose mensual. Años 2-5 con desglose mensual.",
        "▸ Las celdas en verde son campos editables (inputs).",
        "▸ EBITDA y márgenes se calculan automáticamente.",
        "▸ La pestaña Resumen consolida los 5 años.",
    ])

    revenue_lines = ["Comedor", "Barra", "Delivery", "Eventos / Catering"]
    cost_lines = ["Food Cost", "Personal (Labor)", "Alquiler", "Suministros", "Marketing", "Administración / Otros"]

    for year_num in range(1, 6):
        ws = wb.create_sheet(f"Año {year_num}")
        ws.sheet_properties.tabColor = "1565C0"
        set_col_widths(ws, [("A", 28)] + [(get_column_letter(i), 14) for i in range(2, 15)])

        ws.merge_cells("A1:N1")
        ws["A1"].value = f"P&L Previsional — Año {year_num}"
        ws["A1"].font = title_font
        ws.merge_cells("A2:N2")
        ws["A2"].value = BRAND_LINE
        ws["A2"].font = subtitle_font

        headers = ["Concepto"] + MONTHS + ["TOTAL AÑO"]
        style_header_row(ws, 4, headers)

        r = 5
        ws.cell(row=r, column=1, value="INGRESOS").font = section_font
        r += 1
        rev_start = r
        for line in revenue_lines:
            ws.cell(row=r, column=1, value=line).font = data_font
            ws.cell(row=r, column=1).border = thin_border
            for c in range(2, 14):
                style_data_cell(ws, r, c, 0, EUR_FMT, fill=input_fill)
            style_data_cell(ws, r, 14, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
            r += 1
        rev_end = r - 1

        ws.cell(row=r, column=1, value="TOTAL INGRESOS").font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=c).value = f"=SUM({col_l}{rev_start}:{col_l}{rev_end})"
        total_rev_row = r
        r += 2

        ws.cell(row=r, column=1, value="GASTOS").font = section_font
        r += 1
        cost_start = r
        for line in cost_lines:
            ws.cell(row=r, column=1, value=line).font = data_font
            ws.cell(row=r, column=1).border = thin_border
            for c in range(2, 14):
                style_data_cell(ws, r, c, 0, EUR_FMT, fill=input_fill)
            style_data_cell(ws, r, 14, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
            r += 1
        cost_end = r - 1

        ws.cell(row=r, column=1, value="TOTAL GASTOS").font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
            ws.cell(row=r, column=c).value = f"=SUM({col_l}{cost_start}:{col_l}{cost_end})"
        total_cost_row = r
        r += 2

        ws.cell(row=r, column=1, value="EBITDA").font = Font(name="Calibri", size=13, bold=True, color=GOLD)
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=EUR_FMT, font=Font(name="Calibri", size=12, bold=True, color=GOLD))
            ws.cell(row=r, column=c).value = f"={col_l}{total_rev_row}-{col_l}{total_cost_row}"
        ebitda_row = r
        r += 1

        ws.cell(row=r, column=1, value="Margen EBITDA %").font = bold_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 15):
            col_l = get_column_letter(c)
            style_data_cell(ws, r, c, fmt=PCT_FMT, font=pct_font)
            ws.cell(row=r, column=c).value = f'=IF({col_l}{total_rev_row}=0,0,{col_l}{ebitda_row}/{col_l}{total_rev_row})'
        r += 2
        add_brand_footer(ws, r, "N")

    # Resumen 5 años
    ws = wb.create_sheet("Resumen")
    ws.sheet_properties.tabColor = GOLD
    set_col_widths(ws, [("A", 28), ("B", 16), ("C", 16), ("D", 16), ("E", 16), ("F", 16), ("G", 18)])

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Resumen Ejecutivo — P&L 5 Años"
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Concepto", "Año 1", "Año 2", "Año 3", "Año 4", "Año 5", "Crecimiento Y1→Y5"]
    style_header_row(ws, 4, headers)

    summary_items = ["Total Ingresos", "Total Gastos", "EBITDA", "Margen EBITDA %"]
    for i, item in enumerate(summary_items, 5):
        ws.cell(row=i, column=1, value=item).font = bold_font
        ws.cell(row=i, column=1).border = thin_border
        for c in range(2, 8):
            style_data_cell(ws, i, c, 0, EUR_FMT if "%" not in item else PCT_FMT)
    ws.cell(row=5, column=7).value = "=IF(B5=0,0,(F5-B5)/B5)"
    ws.cell(row=5, column=7).number_format = PCT_FMT

    add_brand_footer(ws, 12, "G")

    wb.save(os.path.join(OUTPUT_DIR, "01b-plan-financiero-previsional-5-anos.xlsx"))
    print("  ✓ 01b-plan-financiero-previsional-5-anos.xlsx")


# ═══════════════════════════════════════════════════════════
# 02 — CALCULADORA PUNTO DE EQUILIBRIO
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Calculadora de Punto de Equilibrio", [
        "Cómo usar esta plantilla:",
        "▸ En la pestaña 'Datos', introduce tus costes fijos mensuales, % de coste variable y ticket medio.",
        "▸ La pestaña 'Break-Even' calcula automáticamente la facturación mínima y cubiertos/día necesarios.",
        "▸ La pestaña 'Escenarios' muestra 3 escenarios (pesimista, base, optimista).",
        "",
        "Conceptos clave:",
        "▸ Costes Fijos: alquiler, nóminas fijas, seguros, cuotas — no varían con la actividad.",
        "▸ Costes Variables (%): food cost, extras, comisiones — proporción de cada venta.",
        "▸ Margen de Contribución = 1 - % Coste Variable.",
    ])

    # --- Datos ---
    ws = wb.create_sheet("Datos")
    ws.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws, [("A", 3), ("B", 35), ("C", 20), ("D", 30)])

    ws.merge_cells("B1:D1")
    ws["B1"].value = "Datos de Entrada — Punto de Equilibrio"
    ws["B1"].font = title_font
    ws["B2"].value = BRAND_LINE
    ws["B2"].font = subtitle_font

    style_header_row(ws, 4, ["", "Concepto", "Valor", "Notas"], col_start=1)

    fixed_costs = [
        ("Alquiler mensual", 3000, "€/mes"),
        ("Nóminas fijas (bruto)", 12000, "€/mes — incluye SS empresa"),
        ("Seguros", 400, "€/mes"),
        ("Suministros (luz, agua, gas)", 1500, "€/mes"),
        ("Marketing fijo", 500, "€/mes"),
        ("Gestoría / Administración", 300, "€/mes"),
        ("Cuota préstamo", 800, "€/mes"),
        ("Otros gastos fijos", 500, "€/mes"),
    ]

    r = 5
    ws.cell(row=r, column=2, value="COSTES FIJOS MENSUALES").font = section_font
    r += 1
    for label, val, note in fixed_costs:
        ws.cell(row=r, column=2, value=label).font = data_font
        ws.cell(row=r, column=2).border = thin_border
        style_data_cell(ws, r, 3, val, EUR_FMT, fill=input_fill)
        ws.cell(row=r, column=4, value=note).font = small_font
        r += 1

    total_fixed_row = r
    ws.cell(row=r, column=2, value="TOTAL COSTES FIJOS").font = bold_font
    ws.cell(row=r, column=2).border = thin_border
    style_data_cell(ws, r, 3, fmt=EUR_FMT, font=bold_font)
    ws.cell(row=r, column=3).value = f"=SUM(C6:C{r-1})"
    r += 2

    ws.cell(row=r, column=2, value="COSTES VARIABLES").font = section_font
    r += 1
    var_cost_row = r
    ws.cell(row=r, column=2, value="% Coste Variable sobre ventas").font = data_font
    ws.cell(row=r, column=2).border = thin_border
    style_data_cell(ws, r, 3, 0.35, PCT_FMT, fill=input_fill)
    ws.cell(row=r, column=4, value="Food cost + variable costs (30-40% típico)").font = small_font
    r += 2

    ws.cell(row=r, column=2, value="TICKET MEDIO").font = section_font
    r += 1
    ticket_row = r
    ws.cell(row=r, column=2, value="Ticket medio por comensal").font = data_font
    ws.cell(row=r, column=2).border = thin_border
    style_data_cell(ws, r, 3, 22, EUR_FMT, fill=input_fill)
    r += 1
    days_row = r
    ws.cell(row=r, column=2, value="Días de apertura al mes").font = data_font
    ws.cell(row=r, column=2).border = thin_border
    style_data_cell(ws, r, 3, 26, INT_FMT, fill=input_fill)

    add_brand_footer(ws, r + 3, "D")

    # --- Break-Even ---
    ws2 = wb.create_sheet("Break-Even")
    ws2.sheet_properties.tabColor = GOLD
    set_col_widths(ws2, [("A", 3), ("B", 40), ("C", 22)])

    ws2.merge_cells("B1:C1")
    ws2["B1"].value = "Cálculo del Punto de Equilibrio"
    ws2["B1"].font = title_font
    ws2["B2"].value = BRAND_LINE
    ws2["B2"].font = subtitle_font

    style_header_row(ws2, 4, ["", "Indicador", "Resultado"], col_start=1)

    calcs = [
        ("Total Costes Fijos / mes", f"=Datos!C{total_fixed_row}", EUR_FMT),
        ("% Coste Variable", f"=Datos!C{var_cost_row}", PCT_FMT),
        ("Margen de Contribución", f"=1-Datos!C{var_cost_row}", PCT_FMT),
        ("FACTURACIÓN BREAK-EVEN / mes", f"=Datos!C{total_fixed_row}/(1-Datos!C{var_cost_row})", EUR_FMT),
        ("Facturación Break-Even / día", f"=C8/Datos!C{days_row}", EUR_FMT),
        ("Cubiertos / día necesarios", f"=C9/Datos!C{ticket_row}", INT_FMT),
        ("Cubiertos / mes necesarios", f"=C10*Datos!C{days_row}", INT_FMT),
    ]

    for i, (label, formula, fmt) in enumerate(calcs, 5):
        ws2.cell(row=i, column=2, value=label).font = bold_font if "BREAK-EVEN" in label else data_font
        ws2.cell(row=i, column=2).border = thin_border
        c = style_data_cell(ws2, i, 3, fmt=fmt, font=formula_font if "BREAK-EVEN" in label else data_font)
        c.value = formula

    add_brand_footer(ws2, 15, "C")

    # --- Escenarios ---
    ws3 = wb.create_sheet("Escenarios")
    ws3.sheet_properties.tabColor = "FF6F00"
    set_col_widths(ws3, [("A", 3), ("B", 35), ("C", 18), ("D", 18), ("E", 18)])

    ws3.merge_cells("B1:E1")
    ws3["B1"].value = "Análisis de Escenarios"
    ws3["B1"].font = title_font

    style_header_row(ws3, 3, ["", "Indicador", "Pesimista", "Base", "Optimista"], col_start=1)

    scenarios = [
        ("Ticket medio (€)", 18, 22, 28),
        ("Cubiertos / día", 40, 55, 75),
        ("% Coste Variable", 0.40, 0.35, 0.30),
        ("Costes Fijos / mes (€)", 19000, 19000, 19000),
    ]

    r = 4
    for label, pesi, base, opti in scenarios:
        ws3.cell(row=r, column=2, value=label).font = data_font
        ws3.cell(row=r, column=2).border = thin_border
        fmt = PCT_FMT if "%" in label else (EUR_FMT if "€" in label else INT_FMT)
        style_data_cell(ws3, r, 3, pesi, fmt, fill=rojo_fill)
        style_data_cell(ws3, r, 4, base, fmt, fill=amarillo_fill)
        style_data_cell(ws3, r, 5, opti, fmt, fill=verde_fill)
        r += 1

    r += 1
    ws3.cell(row=r, column=2, value="RESULTADOS").font = section_font
    r += 1
    results = [
        ("Facturación mensual", "={col}4*{col}5*26", EUR_FMT),
        ("Costes Variables", "={col}{fr}*{col}6", EUR_FMT),
        ("Margen Contribución", "={col}{fr}-{col}{fr1}", EUR_FMT),
        ("EBITDA", "={col}{fr2}-{col}7", EUR_FMT),
    ]
    for label, formula_tpl, fmt in results:
        ws3.cell(row=r, column=2, value=label).font = bold_font if "EBITDA" in label else data_font
        ws3.cell(row=r, column=2).border = thin_border
        for ci, col_l in enumerate(["C", "D", "E"]):
            fr = r  # facturacion row
            if "Facturación" in label:
                f = f"={col_l}4*{col_l}5*26"
            elif "Variables" in label:
                f = f"={col_l}{r-1}*{col_l}6"
            elif "Margen" in label:
                f = f"={col_l}{r-2}-{col_l}{r-1}"
            elif "EBITDA" in label:
                f = f"={col_l}{r-1}-{col_l}7"
            else:
                f = 0
            c = style_data_cell(ws3, r, 3 + ci, fmt=fmt, font=bold_font if "EBITDA" in label else formula_font)
            c.value = f
        r += 1

    add_brand_footer(ws3, r + 2, "E")
    wb.save(os.path.join(OUTPUT_DIR, "02-calculadora-punto-equilibrio.xlsx"))
    print("  ✓ 02-calculadora-punto-equilibrio.xlsx")


# ═══════════════════════════════════════════════════════════
# 03 — CASH FLOW FORECAST 12 MESES
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Previsión de Tesorería (Cash Flow) — 12 Meses", [
        "Cómo usar esta plantilla:",
        "▸ Introduce los cobros y pagos previstos mes a mes.",
        "▸ El IVA trimestral se calcula automáticamente (línea 'Liquidación IVA').",
        "▸ La pestaña 'Alertas' marca en rojo los meses donde el saldo < umbral de seguridad.",
        "",
        "Conceptos:",
        "▸ Cobros = ingresos reales de caja (con IVA incluido).",
        "▸ Pagos = salidas reales de caja (con IVA incluido).",
        "▸ Saldo Final = Saldo Inicial + Cobros - Pagos.",
    ])

    ws = wb.create_sheet("Flujo Mensual")
    ws.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws, [("A", 30)] + [(get_column_letter(i), 14) for i in range(2, 14)])

    ws.merge_cells("A1:M1")
    ws["A1"].value = "Cash Flow Forecast — 12 Meses"
    ws["A1"].font = title_font
    ws.merge_cells("A2:M2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Concepto"] + MONTHS
    style_header_row(ws, 4, headers)

    r = 5
    ws.cell(row=r, column=1, value="SALDO INICIAL").font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    style_data_cell(ws, r, 2, 15000, EUR_FMT, fill=input_fill)
    for c in range(3, 14):
        col_l = get_column_letter(c)
        prev_col = get_column_letter(c - 1)
        style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
        # saldo final of prev month
        ws.cell(row=r, column=c).value = f"={prev_col}22"
    saldo_ini_row = r

    r += 1
    ws.cell(row=r, column=1, value="COBROS (Entradas)").font = section_font
    r += 1
    cobros_items = ["Ventas comedor (IVA incl.)", "Ventas barra (IVA incl.)", "Ventas delivery (IVA incl.)",
                    "Eventos / Catering", "Otros cobros"]
    cobros_start = r
    for item in cobros_items:
        ws.cell(row=r, column=1, value=item).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 14):
            style_data_cell(ws, r, c, 0, EUR_FMT, fill=input_fill)
        r += 1
    cobros_end = r - 1
    ws.cell(row=r, column=1, value="TOTAL COBROS").font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    for c in range(2, 14):
        col_l = get_column_letter(c)
        style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
        ws.cell(row=r, column=c).value = f"=SUM({col_l}{cobros_start}:{col_l}{cobros_end})"
    total_cobros_row = r

    r += 1
    ws.cell(row=r, column=1, value="PAGOS (Salidas)").font = section_font
    r += 1
    pagos_items = ["Compras materia prima", "Nóminas + SS", "Alquiler", "Suministros",
                   "Marketing", "Gestoría / Admin", "Cuota préstamo", "Liquidación IVA trimestral",
                   "Otros pagos"]
    pagos_start = r
    for item in pagos_items:
        ws.cell(row=r, column=1, value=item).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 14):
            style_data_cell(ws, r, c, 0, EUR_FMT, fill=input_fill)
        r += 1
    pagos_end = r - 1
    ws.cell(row=r, column=1, value="TOTAL PAGOS").font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    for c in range(2, 14):
        col_l = get_column_letter(c)
        style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
        ws.cell(row=r, column=c).value = f"=SUM({col_l}{pagos_start}:{col_l}{pagos_end})"
    total_pagos_row = r

    r += 1
    # Flujo Neto
    ws.cell(row=r, column=1, value="FLUJO NETO MES").font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    for c in range(2, 14):
        col_l = get_column_letter(c)
        style_data_cell(ws, r, c, fmt=EUR_FMT, font=formula_font)
        ws.cell(row=r, column=c).value = f"={col_l}{total_cobros_row}-{col_l}{total_pagos_row}"
    flujo_row = r

    r += 1
    # Saldo Final
    ws.cell(row=r, column=1, value="SALDO FINAL").font = Font(name="Calibri", size=13, bold=True, color=GOLD)
    ws.cell(row=r, column=1).border = thin_border
    for c in range(2, 14):
        col_l = get_column_letter(c)
        style_data_cell(ws, r, c, fmt=EUR_FMT, font=Font(name="Calibri", size=12, bold=True, color=GOLD))
        ws.cell(row=r, column=c).value = f"={col_l}{saldo_ini_row}+{col_l}{flujo_row}"
    saldo_final_row = r

    r += 2
    add_brand_footer(ws, r, "M")

    # --- Alertas ---
    ws2 = wb.create_sheet("Alertas")
    ws2.sheet_properties.tabColor = "D32F2F"
    set_col_widths(ws2, [("A", 3), ("B", 20), ("C", 18), ("D", 18), ("E", 18)])

    ws2.merge_cells("B1:E1")
    ws2["B1"].value = "Alertas de Tesorería"
    ws2["B1"].font = title_font

    ws2.cell(row=3, column=2, value="Umbral de seguridad (€):").font = bold_font
    style_data_cell(ws2, 3, 3, 5000, EUR_FMT, fill=input_fill)

    style_header_row(ws2, 5, ["", "Mes", "Saldo Final", "Alerta"], col_start=1)

    for i, month in enumerate(MONTHS):
        r = 6 + i
        ws2.cell(row=r, column=2, value=month).font = data_font
        ws2.cell(row=r, column=2).border = thin_border
        col_l = get_column_letter(i + 2)
        c = style_data_cell(ws2, r, 3, fmt=EUR_FMT)
        c.value = f"='Flujo Mensual'!{col_l}{saldo_final_row}"
        alert_cell = ws2.cell(row=r, column=4)
        alert_cell.value = f'=IF(C{r}<$C$3,"⚠️ BAJO MÍNIMO","OK")'
        alert_cell.font = data_font
        alert_cell.border = thin_border
        alert_cell.alignment = center_align

    add_brand_footer(ws2, 20, "E")
    wb.save(os.path.join(OUTPUT_DIR, "03-cash-flow-forecast.xlsx"))
    print("  ✓ 03-cash-flow-forecast.xlsx")


# ═══════════════════════════════════════════════════════════
# 04 — PRESUPUESTO INVERSIÓN / CAPEX
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Presupuesto de Inversión (CAPEX)", [
        "Cómo usar esta plantilla:",
        "▸ Cada pestaña cubre una categoría de inversión inicial.",
        "▸ Introduce el presupuesto estimado y luego el coste real.",
        "▸ La desviación se calcula automáticamente (% sobre presupuesto).",
        "▸ La pestaña Resumen consolida todas las categorías.",
        "",
        "Categorías:",
        "▸ Obra civil, Equipamiento Cocina, Mobiliario Sala, Tecnología, Licencias y Permisos.",
    ])

    categories = {
        "Obra": [
            "Demolición y vaciado", "Albañilería", "Fontanería", "Electricidad",
            "Climatización / Extracción", "Pintura y acabados", "Suelo",
            "Fachada / Rótulo", "Baños clientes", "Vestuarios personal"
        ],
        "Equipamiento Cocina": [
            "Cocina industrial (fogones)", "Horno combinado", "Plancha / Grill",
            "Freidora", "Cámara frigorífica", "Congelador", "Mesa de trabajo inox",
            "Lavavajillas industrial", "Campana extractora", "Pequeño material cocina"
        ],
        "Mobiliario Sala": [
            "Mesas", "Sillas", "Barra", "Taburetes barra", "Estanterías / decoración",
            "Iluminación", "Terraza (mesas + sillas)", "Textil (manteles, servilletas)"
        ],
        "Tecnología": [
            "TPV (hardware + software)", "Impresoras comandas", "Datáfono",
            "Sistema de reservas", "Wifi / Red", "Pantallas / Digital signage",
            "Web / App delivery", "Sistema de seguridad (cámaras)"
        ],
        "Licencias": [
            "Licencia de apertura", "Licencia de obras", "Proyecto técnico (arquitecto)",
            "Certificado energético", "Plan de autoprotección", "Alta Hacienda / SS",
            "Registro sanitario", "SGAE / Derechos música", "Seguros (año 1)"
        ],
    }

    for cat_name, items in categories.items():
        ws = wb.create_sheet(cat_name)
        ws.sheet_properties.tabColor = "1565C0"
        set_col_widths(ws, [("A", 35), ("B", 16), ("C", 16), ("D", 14), ("E", 30)])

        ws.merge_cells("A1:E1")
        ws["A1"].value = f"Inversión — {cat_name}"
        ws["A1"].font = title_font
        ws.merge_cells("A2:E2")
        ws["A2"].value = BRAND_LINE
        ws["A2"].font = subtitle_font

        headers = ["Partida", "Presupuesto (€)", "Real (€)", "Desviación %", "Notas"]
        style_header_row(ws, 4, headers)

        for i, item in enumerate(items, 5):
            ws.cell(row=i, column=1, value=item).font = data_font
            ws.cell(row=i, column=1).border = thin_border
            style_data_cell(ws, i, 2, 0, EUR_FMT, fill=input_fill)
            style_data_cell(ws, i, 3, 0, EUR_FMT, fill=input_fill)
            style_data_cell(ws, i, 4, fmt=PCT_FMT, font=formula_font)
            ws.cell(row=i, column=4).value = f'=IF(B{i}=0,0,(C{i}-B{i})/B{i})'
            ws.cell(row=i, column=5).font = data_font
            ws.cell(row=i, column=5).border = thin_border

        total_row = 5 + len(items)
        ws.cell(row=total_row, column=1, value=f"TOTAL {cat_name.upper()}").font = bold_font
        ws.cell(row=total_row, column=1).border = thin_border
        style_data_cell(ws, total_row, 2, fmt=EUR_FMT, font=bold_font)
        ws.cell(row=total_row, column=2).value = f"=SUM(B5:B{total_row-1})"
        style_data_cell(ws, total_row, 3, fmt=EUR_FMT, font=bold_font)
        ws.cell(row=total_row, column=3).value = f"=SUM(C5:C{total_row-1})"
        style_data_cell(ws, total_row, 4, fmt=PCT_FMT, font=formula_font)
        ws.cell(row=total_row, column=4).value = f'=IF(B{total_row}=0,0,(C{total_row}-B{total_row})/B{total_row})'

        add_brand_footer(ws, total_row + 2, "E")

    # Resumen
    ws = wb.create_sheet("Resumen")
    ws.sheet_properties.tabColor = GOLD
    set_col_widths(ws, [("A", 30), ("B", 18), ("C", 18), ("D", 16)])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Resumen de Inversión Total"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Categoría", "Presupuesto (€)", "Real (€)", "Desviación %"]
    style_header_row(ws, 4, headers)

    cat_names = list(categories.keys())
    for i, cat in enumerate(cat_names, 5):
        total_row_num = 5 + len(categories[cat])
        ws.cell(row=i, column=1, value=cat).font = data_font
        ws.cell(row=i, column=1).border = thin_border
        style_data_cell(ws, i, 2, fmt=EUR_FMT)
        ws.cell(row=i, column=2).value = f"='{cat}'!B{total_row_num}"
        style_data_cell(ws, i, 3, fmt=EUR_FMT)
        ws.cell(row=i, column=3).value = f"='{cat}'!C{total_row_num}"
        style_data_cell(ws, i, 4, fmt=PCT_FMT, font=formula_font)
        ws.cell(row=i, column=4).value = f'=IF(B{i}=0,0,(C{i}-B{i})/B{i})'

    total_r = 5 + len(cat_names)
    ws.cell(row=total_r, column=1, value="INVERSIÓN TOTAL").font = Font(name="Calibri", size=13, bold=True, color=GOLD)
    ws.cell(row=total_r, column=1).border = thin_border
    style_data_cell(ws, total_r, 2, fmt=EUR_FMT, font=Font(name="Calibri", size=12, bold=True, color=GOLD))
    ws.cell(row=total_r, column=2).value = f"=SUM(B5:B{total_r-1})"
    style_data_cell(ws, total_r, 3, fmt=EUR_FMT, font=Font(name="Calibri", size=12, bold=True, color=GOLD))
    ws.cell(row=total_r, column=3).value = f"=SUM(C5:C{total_r-1})"
    style_data_cell(ws, total_r, 4, fmt=PCT_FMT, font=formula_font)
    ws.cell(row=total_r, column=4).value = f'=IF(B{total_r}=0,0,(C{total_r}-B{total_r})/B{total_r})'

    add_brand_footer(ws, total_r + 3, "D")
    wb.save(os.path.join(OUTPUT_DIR, "04-presupuesto-inversion-capex.xlsx"))
    print("  ✓ 04-presupuesto-inversion-capex.xlsx")


# ═══════════════════════════════════════════════════════════
# 05 — P&L MENSUAL REAL VS PRESUPUESTO
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · P&L Mensual — Real vs Presupuesto", [
        "Cómo usar esta plantilla:",
        "▸ Cada mes tiene su pestaña con columnas: Presupuesto, Real, Desviación €, Desviación %.",
        "▸ El semáforo automático marca: Verde (<5%), Amarillo (5-10%), Rojo (>10%).",
        "▸ Food Cost %, Labor Cost % y Prime Cost % se calculan automáticamente.",
        "▸ La pestaña Resumen consolida los 12 meses.",
    ])

    lines = [
        ("INGRESOS", None),
        ("Ventas Comedor", "rev"),
        ("Ventas Barra", "rev"),
        ("Ventas Delivery", "rev"),
        ("Eventos / Catering", "rev"),
        ("TOTAL INGRESOS", "total_rev"),
        ("", None),
        ("GASTOS", None),
        ("Food Cost", "cost"),
        ("Personal (Labor)", "cost"),
        ("Alquiler", "cost"),
        ("Suministros", "cost"),
        ("Marketing", "cost"),
        ("Administración / Otros", "cost"),
        ("TOTAL GASTOS", "total_cost"),
        ("", None),
        ("EBITDA", "ebitda"),
        ("Margen EBITDA %", "pct"),
        ("", None),
        ("RATIOS AUTOMÁTICOS", None),
        ("Food Cost %", "ratio"),
        ("Labor Cost %", "ratio"),
        ("Prime Cost %", "ratio"),
    ]

    for mi, month in enumerate(MONTHS):
        ws = wb.create_sheet(month)
        ws.sheet_properties.tabColor = "1565C0"
        set_col_widths(ws, [("A", 28), ("B", 16), ("C", 16), ("D", 16), ("E", 14), ("F", 14)])

        ws.merge_cells("A1:F1")
        ws["A1"].value = f"P&L {month} — Real vs Presupuesto"
        ws["A1"].font = title_font
        ws.merge_cells("A2:F2")
        ws["A2"].value = BRAND_LINE
        ws["A2"].font = subtitle_font

        headers = ["Concepto", "Presupuesto (€)", "Real (€)", "Desviación (€)", "Desviación %", "Semáforo"]
        style_header_row(ws, 4, headers)

        rev_rows = []
        cost_rows = []
        total_rev_row = None
        total_cost_row = None
        ebitda_row_num = None
        food_cost_row = None
        labor_cost_row = None

        r = 5
        for label, typ in lines:
            if label == "":
                r += 1
                continue
            if typ is None:
                ws.cell(row=r, column=1, value=label).font = section_font
                r += 1
                continue

            ws.cell(row=r, column=1, value=label).font = bold_font if "TOTAL" in label or "EBITDA" in label else data_font
            ws.cell(row=r, column=1).border = thin_border

            if typ == "rev":
                rev_rows.append(r)
                style_data_cell(ws, r, 2, 0, EUR_FMT, fill=input_fill)
                style_data_cell(ws, r, 3, 0, EUR_FMT, fill=input_fill)
                if label == "Food Cost":
                    food_cost_row = r
            elif typ == "cost":
                cost_rows.append(r)
                style_data_cell(ws, r, 2, 0, EUR_FMT, fill=input_fill)
                style_data_cell(ws, r, 3, 0, EUR_FMT, fill=input_fill)
                if "Food Cost" in label:
                    food_cost_row = r
                if "Personal" in label:
                    labor_cost_row = r
            elif typ == "total_rev":
                total_rev_row = r
                for c in [2, 3]:
                    col_l = get_column_letter(c)
                    style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
                    cells_str = "+".join([f"{col_l}{rr}" for rr in rev_rows])
                    ws.cell(row=r, column=c).value = f"={cells_str}"
            elif typ == "total_cost":
                total_cost_row = r
                for c in [2, 3]:
                    col_l = get_column_letter(c)
                    style_data_cell(ws, r, c, fmt=EUR_FMT, font=bold_font)
                    cells_str = "+".join([f"{col_l}{rr}" for rr in cost_rows])
                    ws.cell(row=r, column=c).value = f"={cells_str}"
            elif typ == "ebitda":
                ebitda_row_num = r
                for c in [2, 3]:
                    col_l = get_column_letter(c)
                    style_data_cell(ws, r, c, fmt=EUR_FMT, font=Font(name="Calibri", size=12, bold=True, color=GOLD))
                    ws.cell(row=r, column=c).value = f"={col_l}{total_rev_row}-{col_l}{total_cost_row}"
            elif typ == "pct":
                for c in [2, 3]:
                    col_l = get_column_letter(c)
                    style_data_cell(ws, r, c, fmt=PCT_FMT, font=pct_font)
                    ws.cell(row=r, column=c).value = f"=IF({col_l}{total_rev_row}=0,0,{col_l}{ebitda_row_num}/{col_l}{total_rev_row})"
            elif typ == "ratio":
                if "Food" in label:
                    food_cost_row_ref = cost_rows[0] if cost_rows else r
                    for c in [2, 3]:
                        col_l = get_column_letter(c)
                        style_data_cell(ws, r, c, fmt=PCT_FMT, font=pct_font)
                        ws.cell(row=r, column=c).value = f"=IF({col_l}{total_rev_row}=0,0,{col_l}{food_cost_row_ref}/{col_l}{total_rev_row})"
                elif "Labor" in label:
                    labor_row_ref = cost_rows[1] if len(cost_rows) > 1 else r
                    for c in [2, 3]:
                        col_l = get_column_letter(c)
                        style_data_cell(ws, r, c, fmt=PCT_FMT, font=pct_font)
                        ws.cell(row=r, column=c).value = f"=IF({col_l}{total_rev_row}=0,0,{col_l}{labor_row_ref}/{col_l}{total_rev_row})"
                elif "Prime" in label:
                    fc_ref = cost_rows[0] if cost_rows else r
                    lc_ref = cost_rows[1] if len(cost_rows) > 1 else r
                    for c in [2, 3]:
                        col_l = get_column_letter(c)
                        style_data_cell(ws, r, c, fmt=PCT_FMT, font=pct_font)
                        ws.cell(row=r, column=c).value = f"=IF({col_l}{total_rev_row}=0,0,({col_l}{fc_ref}+{col_l}{lc_ref})/{col_l}{total_rev_row})"

            # Deviation columns for data rows
            if typ in ("rev", "cost", "total_rev", "total_cost", "ebitda"):
                style_data_cell(ws, r, 4, fmt=EUR_FMT, font=formula_font)
                ws.cell(row=r, column=4).value = f"=C{r}-B{r}"
                style_data_cell(ws, r, 5, fmt=PCT_FMT, font=formula_font)
                ws.cell(row=r, column=5).value = f"=IF(B{r}=0,0,(C{r}-B{r})/B{r})"
                # Semaphore
                ws.cell(row=r, column=6).value = f'=IF(ABS(E{r})<0.05,"✅ OK",IF(ABS(E{r})<0.1,"⚠️ Atención","🔴 Alerta"))'
                ws.cell(row=r, column=6).font = data_font
                ws.cell(row=r, column=6).border = thin_border
                ws.cell(row=r, column=6).alignment = center_align

            r += 1

        add_brand_footer(ws, r + 1, "F")

    # Resumen tab
    ws = wb.create_sheet("Resumen Anual")
    ws.sheet_properties.tabColor = GOLD
    set_col_widths(ws, [("A", 28)] + [(get_column_letter(i), 14) for i in range(2, 15)])

    ws.merge_cells("A1:N1")
    ws["A1"].value = "Resumen Anual — Real vs Presupuesto"
    ws["A1"].font = title_font
    ws.merge_cells("A2:N2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["Concepto"] + MONTHS + ["TOTAL"]
    style_header_row(ws, 4, headers)

    summary_rows = ["Total Ingresos (Ppto)", "Total Ingresos (Real)",
                    "Total Gastos (Ppto)", "Total Gastos (Real)",
                    "EBITDA (Ppto)", "EBITDA (Real)"]
    for i, label in enumerate(summary_rows, 5):
        ws.cell(row=i, column=1, value=label).font = bold_font
        ws.cell(row=i, column=1).border = thin_border
        for c in range(2, 15):
            style_data_cell(ws, i, c, 0, EUR_FMT)

    add_brand_footer(ws, 14, "N")
    wb.save(os.path.join(OUTPUT_DIR, "05-pyl-mensual-real-vs-presupuesto.xlsx"))
    print("  ✓ 05-pyl-mensual-real-vs-presupuesto.xlsx")


# ═══════════════════════════════════════════════════════════
# 06 — DASHBOARD RATIOS FINANCIEROS
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Dashboard de Ratios Financieros", [
        "Cómo usar esta plantilla:",
        "▸ Introduce tus datos reales en la columna 'Tu Restaurante'.",
        "▸ Los benchmarks del sector aparecen para comparar tu rendimiento.",
        "▸ El semáforo indica si estás dentro de rango óptimo (verde), aceptable (amarillo) o peligro (rojo).",
        "",
        "Ratios clave hostelería:",
        "▸ Food Cost ideal: 28-32% sobre ventas.",
        "▸ Labor Cost ideal: 25-30% sobre ventas.",
        "▸ Prime Cost (Food + Labor): máximo 60-65%.",
        "▸ GOP (Gross Operating Profit): objetivo >15%.",
    ])

    ws = wb.create_sheet("Ratios")
    ws.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws, [("A", 3), ("B", 35), ("C", 18), ("D", 20), ("E", 18)])

    ws.merge_cells("B1:E1")
    ws["B1"].value = "Dashboard de Ratios Financieros"
    ws["B1"].font = title_font
    ws.merge_cells("B2:E2")
    ws["B2"].value = BRAND_LINE
    ws["B2"].font = subtitle_font

    # Input section
    ws.cell(row=4, column=2, value="DATOS DE ENTRADA").font = section_font
    style_header_row(ws, 5, ["", "Concepto", "Valor", "Período", ""], col_start=1)

    inputs = [
        ("Ventas totales (€)", 85000, "Mensual"),
        ("Coste materia prima (€)", 25500, "Mensual"),
        ("Coste personal total (€)", 23800, "Mensual"),
        ("Gastos operativos totales (€)", 72250, "Mensual"),
        ("Nº de cubiertos servidos", 3200, "Mensual"),
        ("m² de sala", 80, ""),
        ("Horas de servicio / mes", 360, ""),
    ]

    for i, (label, val, period) in enumerate(inputs, 6):
        ws.cell(row=i, column=2, value=label).font = data_font
        ws.cell(row=i, column=2).border = thin_border
        fmt = EUR_FMT if "€" in label else INT_FMT
        style_data_cell(ws, i, 3, val, fmt, fill=input_fill)
        ws.cell(row=i, column=4, value=period).font = small_font

    r = 15
    ws.cell(row=r, column=2, value="RATIOS CALCULADOS").font = section_font
    style_header_row(ws, r + 1, ["", "Ratio", "Tu Valor", "Benchmark Sector", "Estado"], col_start=1)

    ratios = [
        ("Food Cost %", "=C8/C6", "28% - 32%"),
        ("Labor Cost %", "=C8/C6", "25% - 30%"),
        ("Prime Cost % (Food + Labor)", "=(C7+C8)/C6", "< 65%"),
        ("GOP (Gross Operating Profit) %", "=(C6-C9)/C6", "> 15%"),
        ("Coste por cubierto (€)", "=C9/C10", "Varía por concepto"),
        ("RevPASH (€/m²/hora)", "=C6/(C11*C12)", "> 8€ casual dining"),
        ("Ticket medio (€)", "=C6/C10", "Varía por concepto"),
        ("Margen bruto %", "=(C6-C7)/C6", "> 65%"),
    ]

    # Fix formulas for correct references
    ratio_formulas = [
        "=C7/C6",   # Food Cost
        "=C8/C6",   # Labor Cost
        "=(C7+C8)/C6",  # Prime Cost
        "=(C6-C9)/C6",  # GOP
        "=C9/C10",  # Coste por cubierto
        "=C6/(C11*C12)",  # RevPASH
        "=C6/C10",  # Ticket medio
        "=(C6-C7)/C6",  # Margen bruto
    ]

    for i, (label, _, benchmark) in enumerate(ratios):
        row = r + 2 + i
        ws.cell(row=row, column=2, value=label).font = bold_font
        ws.cell(row=row, column=2).border = thin_border
        fmt = PCT_FMT if "%" in label else EUR_FMT
        c = style_data_cell(ws, row, 3, fmt=fmt, font=formula_font)
        c.value = ratio_formulas[i]
        ws.cell(row=row, column=4, value=benchmark).font = data_font
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).alignment = center_align

    # Semaphore for food cost
    fc_row = r + 2
    ws.cell(row=fc_row, column=5).value = f'=IF(C{fc_row}<0.28,"✅ Excelente",IF(C{fc_row}<0.33,"⚠️ Aceptable","🔴 Alto"))'
    lc_row = r + 3
    ws.cell(row=lc_row, column=5).value = f'=IF(C{lc_row}<0.25,"✅ Excelente",IF(C{lc_row}<0.31,"⚠️ Aceptable","🔴 Alto"))'
    pc_row = r + 4
    ws.cell(row=pc_row, column=5).value = f'=IF(C{pc_row}<0.6,"✅ Excelente",IF(C{pc_row}<0.66,"⚠️ Aceptable","🔴 Alto"))'
    gop_row = r + 5
    ws.cell(row=gop_row, column=5).value = f'=IF(C{gop_row}>0.15,"✅ Sano",IF(C{gop_row}>0.1,"⚠️ Ajustado","🔴 Peligro"))'

    add_brand_footer(ws, r + 12, "E")

    # --- Benchmarks ---
    ws2 = wb.create_sheet("Benchmarks")
    ws2.sheet_properties.tabColor = GOLD
    set_col_widths(ws2, [("A", 3), ("B", 35), ("C", 18), ("D", 18), ("E", 18)])

    ws2.merge_cells("B1:E1")
    ws2["B1"].value = "Benchmarks del Sector Hostelería"
    ws2["B1"].font = title_font

    style_header_row(ws2, 3, ["", "Ratio", "Óptimo", "Aceptable", "Peligro"], col_start=1)

    benchmarks = [
        ("Food Cost %", "< 28%", "28% - 33%", "> 33%"),
        ("Labor Cost %", "< 25%", "25% - 30%", "> 30%"),
        ("Prime Cost %", "< 60%", "60% - 65%", "> 65%"),
        ("GOP %", "> 20%", "15% - 20%", "< 15%"),
        ("Alquiler / Ventas", "< 8%", "8% - 12%", "> 12%"),
        ("Marketing / Ventas", "3% - 5%", "5% - 8%", "> 8%"),
        ("EBITDA %", "> 15%", "10% - 15%", "< 10%"),
        ("Coste cubierto / Ticket", "< 60%", "60% - 70%", "> 70%"),
    ]

    for i, (label, opt, acc, danger) in enumerate(benchmarks, 4):
        ws2.cell(row=i, column=2, value=label).font = data_font
        ws2.cell(row=i, column=2).border = thin_border
        style_data_cell(ws2, i, 3, opt, font=data_font, fill=verde_fill, align=center_align)
        style_data_cell(ws2, i, 4, acc, font=data_font, fill=amarillo_fill, align=center_align)
        style_data_cell(ws2, i, 5, danger, font=data_font, fill=rojo_fill, align=center_align)

    add_brand_footer(ws2, 14, "E")
    wb.save(os.path.join(OUTPUT_DIR, "06-dashboard-ratios-financieros.xlsx"))
    print("  ✓ 06-dashboard-ratios-financieros.xlsx")


# ═══════════════════════════════════════════════════════════
# 07 — INFORME DE VIABILIDAD PARA BANCOS
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Informe de Viabilidad — Presentación Bancaria", [
        "Cómo usar esta plantilla:",
        "▸ Completa cada pestaña con los datos de tu proyecto.",
        "▸ El formato profesional está listo para presentar directamente al banco.",
        "▸ TIR, VAN y Payback se calculan automáticamente.",
        "▸ Incluye sección de garantías y avales.",
        "",
        "Pestañas:",
        "▸ Resumen Ejecutivo: datos clave del proyecto.",
        "▸ Proyecciones: P&L + Cash Flow resumido 5 años.",
        "▸ Ratios: indicadores de solvencia y rentabilidad.",
        "▸ Garantías: avales y garantías ofrecidas.",
    ])

    # --- Resumen Ejecutivo ---
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws, [("A", 3), ("B", 35), ("C", 40)])

    ws.merge_cells("B1:C1")
    ws["B1"].value = "Resumen Ejecutivo del Proyecto"
    ws["B1"].font = title_font
    ws["B2"].value = BRAND_LINE
    ws["B2"].font = subtitle_font

    exec_items = [
        ("DATOS DEL PROYECTO", None),
        ("Nombre del restaurante", ""),
        ("Concepto / Tipo de cocina", ""),
        ("Ubicación", ""),
        ("Superficie total (m²)", ""),
        ("Aforo (cubiertos)", ""),
        ("Fecha prevista de apertura", ""),
        ("", None),
        ("DATOS FINANCIEROS CLAVE", None),
        ("Inversión total necesaria", ""),
        ("Fondos propios aportados", ""),
        ("Financiación solicitada", ""),
        ("Plazo de amortización solicitado", ""),
        ("Facturación prevista Año 1", ""),
        ("EBITDA previsto Año 1", ""),
        ("Payback estimado", ""),
        ("", None),
        ("EQUIPO PROMOTOR", None),
        ("Nombre del promotor/a", ""),
        ("Experiencia en hostelería", ""),
        ("Formación", ""),
        ("Otros socios", ""),
    ]

    r = 4
    for label, default in exec_items:
        if label == "":
            r += 1
            continue
        if default is None:
            ws.cell(row=r, column=2, value=label).font = section_font
            r += 1
            continue
        ws.cell(row=r, column=2, value=label).font = data_font
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=3).fill = input_fill
        ws.cell(row=r, column=3).font = data_font
        r += 1

    add_brand_footer(ws, r + 1, "C")

    # --- Proyecciones ---
    ws2 = wb.create_sheet("Proyecciones")
    ws2.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws2, [("A", 30)] + [(get_column_letter(i), 16) for i in range(2, 7)])

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "Proyecciones Financieras — 5 Años"
    ws2["A1"].font = title_font

    headers = ["Concepto", "Año 1", "Año 2", "Año 3", "Año 4", "Año 5"]
    style_header_row(ws2, 3, headers)

    proj_lines = [
        "Facturación", "Food Cost", "Personal", "Otros Gastos Operativos",
        "TOTAL GASTOS", "EBITDA", "Amortización", "Intereses deuda",
        "BAI (Beneficio Antes Impuestos)", "Impuesto Sociedades (25%)",
        "BENEFICIO NETO", "", "CASH FLOW OPERATIVO",
    ]

    for i, line in enumerate(proj_lines, 4):
        if line == "":
            continue
        ws2.cell(row=i, column=1, value=line).font = bold_font if line.isupper() or "NETO" in line or "CASH" in line else data_font
        ws2.cell(row=i, column=1).border = thin_border
        for c in range(2, 7):
            style_data_cell(ws2, i, c, 0, EUR_FMT, fill=input_fill if i < 8 else None)

    r = 4 + len(proj_lines) + 1
    # TIR, VAN, Payback
    ws2.cell(row=r, column=1, value="INDICADORES DE RENTABILIDAD").font = section_font
    r += 1

    ws2.cell(row=r, column=1, value="Inversión inicial (negativa)").font = data_font
    ws2.cell(row=r, column=1).border = thin_border
    style_data_cell(ws2, r, 2, -150000, EUR_FMT, fill=input_fill)
    inv_row = r
    r += 1

    ws2.cell(row=r, column=1, value="Flujos de caja anuales").font = data_font
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2, value="→ Ver fila Cash Flow Operativo arriba").font = small_font
    r += 1

    ws2.cell(row=r, column=1, value="TIR (Tasa Interna de Retorno)").font = bold_font
    ws2.cell(row=r, column=1).border = thin_border
    style_data_cell(ws2, r, 2, fmt=PCT_FMT, font=Font(name="Calibri", size=13, bold=True, color=GOLD))
    # TIR formula using cash flow row (row 16)
    ws2.cell(row=r, column=2).value = f"=IFERROR(IRR(B{inv_row}:F{inv_row})," + '"N/A")'
    r += 1

    ws2.cell(row=r, column=1, value="VAN (al 8% descuento)").font = bold_font
    ws2.cell(row=r, column=1).border = thin_border
    style_data_cell(ws2, r, 2, fmt=EUR_FMT, font=formula_font)
    ws2.cell(row=r, column=2).value = f"=IFERROR(NPV(0.08,C{inv_row}:F{inv_row})+B{inv_row}," + '"N/A")'
    r += 1

    ws2.cell(row=r, column=1, value="Payback (años)").font = bold_font
    ws2.cell(row=r, column=1).border = thin_border
    style_data_cell(ws2, r, 2, fmt='0.0', fill=input_fill)

    add_brand_footer(ws2, r + 3, "F")

    # --- Ratios ---
    ws3 = wb.create_sheet("Ratios")
    ws3.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws3, [("A", 3), ("B", 35), ("C", 18), ("D", 25)])

    ws3.merge_cells("B1:D1")
    ws3["B1"].value = "Ratios de Solvencia y Rentabilidad"
    ws3["B1"].font = title_font

    style_header_row(ws3, 3, ["", "Ratio", "Valor", "Referencia Bancaria"], col_start=1)

    bank_ratios = [
        ("Ratio de endeudamiento", "", "< 60%"),
        ("Cobertura de deuda (DSCR)", "", "> 1.25x"),
        ("Ratio de liquidez corriente", "", "> 1.0"),
        ("Margen EBITDA / Ventas", "", "> 12%"),
        ("ROI (Return on Investment)", "", "> 10%"),
        ("Fondos propios / Inversión", "", "> 30%"),
        ("Deuda / EBITDA", "", "< 3.5x"),
    ]

    for i, (label, val, ref) in enumerate(bank_ratios, 4):
        ws3.cell(row=i, column=2, value=label).font = data_font
        ws3.cell(row=i, column=2).border = thin_border
        style_data_cell(ws3, i, 3, val, fill=input_fill, align=center_align)
        ws3.cell(row=i, column=4, value=ref).font = data_font
        ws3.cell(row=i, column=4).border = thin_border
        ws3.cell(row=i, column=4).alignment = center_align

    add_brand_footer(ws3, 13, "D")

    # --- Garantías ---
    ws4 = wb.create_sheet("Garantías")
    ws4.sheet_properties.tabColor = "FF6F00"
    set_col_widths(ws4, [("A", 3), ("B", 30), ("C", 20), ("D", 35)])

    ws4.merge_cells("B1:D1")
    ws4["B1"].value = "Garantías y Avales Ofrecidos"
    ws4["B1"].font = title_font

    style_header_row(ws4, 3, ["", "Tipo de Garantía", "Valor (€)", "Descripción"], col_start=1)

    guarantees = [
        "Aval personal del promotor",
        "Hipoteca sobre local (si propiedad)",
        "Pignoración de depósitos",
        "Aval SGR (Sociedad de Garantía Recíproca)",
        "Seguro de caución",
        "Fianza / depósito adicional",
    ]

    for i, g in enumerate(guarantees, 4):
        ws4.cell(row=i, column=2, value=g).font = data_font
        ws4.cell(row=i, column=2).border = thin_border
        style_data_cell(ws4, i, 3, 0, EUR_FMT, fill=input_fill)
        ws4.cell(row=i, column=4).border = thin_border
        ws4.cell(row=i, column=4).fill = input_fill
        ws4.cell(row=i, column=4).font = data_font

    total_r = 4 + len(guarantees)
    ws4.cell(row=total_r, column=2, value="TOTAL GARANTÍAS").font = bold_font
    ws4.cell(row=total_r, column=2).border = thin_border
    style_data_cell(ws4, total_r, 3, fmt=EUR_FMT, font=bold_font)
    ws4.cell(row=total_r, column=3).value = f"=SUM(C4:C{total_r-1})"

    add_brand_footer(ws4, total_r + 3, "D")
    wb.save(os.path.join(OUTPUT_DIR, "07-informe-viabilidad-bancos.xlsx"))
    print("  ✓ 07-informe-viabilidad-bancos.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS 08 — SIMULADOR DE ESCENARIOS
# ═══════════════════════════════════════════════════════════
def gen_08():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 08 · Simulador What-If de Escenarios", [
        "Cómo usar esta plantilla:",
        "▸ Cambia los valores de Ticket Medio, Ocupación y Food Cost en la zona de inputs.",
        "▸ El P&L se recalcula instantáneamente para cada escenario.",
        "▸ Compara 3 escenarios lado a lado: Pesimista, Base, Optimista.",
        "",
        "Ideal para:",
        "▸ Preparar planes de contingencia.",
        "▸ Negociar con inversores o socios.",
        "▸ Evaluar el impacto de subidas de precios o cambios de carta.",
    ])

    ws = wb.create_sheet("Simulador")
    ws.sheet_properties.tabColor = "FF6F00"
    set_col_widths(ws, [("A", 30), ("B", 18), ("C", 18), ("D", 18)])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Simulador de Escenarios — What If"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    style_header_row(ws, 4, ["Variable / Resultado", "Pesimista", "Base", "Optimista"])

    # INPUTS
    r = 5
    ws.cell(row=r, column=1, value="VARIABLES DE ENTRADA").font = section_font
    r += 1

    inputs = [
        ("Ticket medio (€)", 16, 22, 30),
        ("Cubiertos / día", 40, 60, 85),
        ("Días apertura / mes", 25, 26, 27),
        ("Food Cost %", 0.38, 0.32, 0.28),
        ("Labor Cost %", 0.32, 0.28, 0.25),
        ("Costes fijos mensuales (€)", 19500, 19000, 18500),
    ]

    input_rows = {}
    for label, pesi, base, opti in inputs:
        ws.cell(row=r, column=1, value=label).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        fmt = PCT_FMT if "%" in label else (EUR_FMT if "€" in label else INT_FMT)
        style_data_cell(ws, r, 2, pesi, fmt, fill=input_fill)
        style_data_cell(ws, r, 3, base, fmt, fill=input_fill)
        style_data_cell(ws, r, 4, opti, fmt, fill=input_fill)
        input_rows[label] = r
        r += 1

    tr = input_rows["Ticket medio (€)"]
    cr = input_rows["Cubiertos / día"]
    dr = input_rows["Días apertura / mes"]
    fcr = input_rows["Food Cost %"]
    lcr = input_rows["Labor Cost %"]
    cfr = input_rows["Costes fijos mensuales (€)"]

    r += 1
    ws.cell(row=r, column=1, value="P&L MENSUAL SIMULADO").font = section_font
    r += 1

    results = [
        ("Facturación mensual", f"={{col}}{tr}*{{col}}{cr}*{{col}}{dr}", EUR_FMT),
        ("Food Cost (€)", f"={{col}}{{}}*{{col}}{fcr}", EUR_FMT),  # needs facturacion row
        ("Labor Cost (€)", f"={{col}}{{}}*{{col}}{lcr}", EUR_FMT),
        ("Costes fijos (€)", f"={{col}}{cfr}", EUR_FMT),
        ("TOTAL GASTOS", None, EUR_FMT),
        ("EBITDA MENSUAL", None, EUR_FMT),
        ("Margen EBITDA %", None, PCT_FMT),
        ("EBITDA ANUAL", None, EUR_FMT),
    ]

    fact_row = r
    for i, (label, _, fmt) in enumerate(results):
        row = r + i
        ws.cell(row=row, column=1, value=label).font = bold_font if "EBITDA" in label or "TOTAL" in label else data_font
        ws.cell(row=row, column=1).border = thin_border

        for ci, col_l in enumerate(["B", "C", "D"]):
            if label == "Facturación mensual":
                f = f"={col_l}{tr}*{col_l}{cr}*{col_l}{dr}"
            elif label == "Food Cost (€)":
                f = f"={col_l}{fact_row}*{col_l}{fcr}"
            elif label == "Labor Cost (€)":
                f = f"={col_l}{fact_row}*{col_l}{lcr}"
            elif label == "Costes fijos (€)":
                f = f"={col_l}{cfr}"
            elif label == "TOTAL GASTOS":
                f = f"={col_l}{fact_row+1}+{col_l}{fact_row+2}+{col_l}{fact_row+3}"
            elif label == "EBITDA MENSUAL":
                f = f"={col_l}{fact_row}-{col_l}{fact_row+4}"
            elif label == "Margen EBITDA %":
                f = f"=IF({col_l}{fact_row}=0,0,{col_l}{fact_row+5}/{col_l}{fact_row})"
            elif label == "EBITDA ANUAL":
                f = f"={col_l}{fact_row+5}*12"
            else:
                f = 0

            font_to_use = Font(name="Calibri", size=12, bold=True, color=GOLD) if "EBITDA" in label and "%" not in label else formula_font
            c = style_data_cell(ws, row, 2 + ci, fmt=fmt, font=font_to_use)
            c.value = f

    add_brand_footer(ws, r + len(results) + 2, "D")

    # --- Comparativa ---
    ws2 = wb.create_sheet("Comparativa")
    ws2.sheet_properties.tabColor = GOLD
    set_col_widths(ws2, [("A", 3), ("B", 30), ("C", 18), ("D", 18), ("E", 18)])

    ws2.merge_cells("B1:E1")
    ws2["B1"].value = "Comparativa de Escenarios — Resumen Anual"
    ws2["B1"].font = title_font

    style_header_row(ws2, 3, ["", "Indicador", "Pesimista", "Base", "Optimista"], col_start=1)

    annual_items = [
        ("Facturación anual", EUR_FMT),
        ("Food Cost anual", EUR_FMT),
        ("Labor Cost anual", EUR_FMT),
        ("EBITDA anual", EUR_FMT),
        ("Margen EBITDA %", PCT_FMT),
    ]

    for i, (label, fmt) in enumerate(annual_items, 4):
        ws2.cell(row=i, column=2, value=label).font = bold_font if "EBITDA" in label else data_font
        ws2.cell(row=i, column=2).border = thin_border
        for c in [3, 4, 5]:
            col_l = get_column_letter(c - 1)  # B, C, D in Simulador
            if "Facturación" in label:
                f = f"=Simulador!{col_l}{fact_row}*12"
            elif "Food" in label:
                f = f"=Simulador!{col_l}{fact_row+1}*12"
            elif "Labor" in label:
                f = f"=Simulador!{col_l}{fact_row+2}*12"
            elif "EBITDA anual" in label:
                f = f"=Simulador!{col_l}{fact_row+7}"
            elif "Margen" in label:
                f = f"=Simulador!{col_l}{fact_row+6}"
            else:
                f = 0
            cell = style_data_cell(ws2, i, c, fmt=fmt, font=formula_font)
            cell.value = f

    add_brand_footer(ws2, 11, "E")
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-08-simulador-escenarios.xlsx"))
    print("  ✓ BONUS-08-simulador-escenarios.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS 09 — CHECKLIST PRE-APERTURA FINANCIERO (48 items)
# ═══════════════════════════════════════════════════════════
def gen_09():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 09 · Checklist Financiero Pre-Apertura — 48 Tareas", [
        "Cómo usar esta plantilla:",
        "▸ 48 tareas organizadas en 6 fases: Constitución, Financiación, Licencias, Proveedores, Seguros, Tesorería.",
        "▸ Marca el estado de cada tarea: Pendiente, En curso, Completada.",
        "▸ Asigna responsable y fecha límite.",
        "▸ Usa la columna Notas para detalles adicionales.",
        "",
        "Fases:",
        "▸ Fase 1: Constitución y estructura legal.",
        "▸ Fase 2: Financiación y captación de fondos.",
        "▸ Fase 3: Licencias, permisos y trámites.",
        "▸ Fase 4: Proveedores y contratos.",
        "▸ Fase 5: Seguros obligatorios y recomendados.",
        "▸ Fase 6: Tesorería y sistemas de cobro.",
    ])

    ws = wb.create_sheet("Checklist")
    ws.sheet_properties.tabColor = "1565C0"
    set_col_widths(ws, [("A", 5), ("B", 22), ("C", 45), ("D", 18), ("E", 15), ("F", 14), ("G", 30)])

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Checklist Financiero Pre-Apertura — 48 Tareas"
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"].value = BRAND_LINE
    ws["A2"].font = subtitle_font

    headers = ["#", "Fase", "Tarea", "Responsable", "Fecha Límite", "Estado", "Notas"]
    style_header_row(ws, 4, headers)

    # Data validation for Estado
    dv = DataValidation(type="list", formula1='"Pendiente,En curso,Completada"', allow_blank=True)
    dv.error = "Selecciona: Pendiente, En curso o Completada"
    dv.errorTitle = "Estado no válido"
    ws.add_data_validation(dv)

    tasks = {
        "1. Constitución": [
            "Elegir forma jurídica (SL, autónomo, cooperativa)",
            "Escritura de constitución ante notario",
            "Inscripción en Registro Mercantil",
            "Obtener NIF/CIF definitivo",
            "Apertura de cuenta bancaria empresarial",
            "Libro de actas y libro de socios",
            "Alta en el Censo de Empresarios (Modelo 036/037)",
            "Registro de marca / nombre comercial (OEPM)",
        ],
        "2. Financiación": [
            "Elaborar plan financiero completo",
            "Calcular necesidades de inversión (CAPEX)",
            "Determinar capital circulante necesario (3-6 meses)",
            "Solicitar préstamo ICO / línea de crédito",
            "Evaluar subvenciones autonómicas/locales",
            "Presentar proyecto a inversores privados",
            "Negociar condiciones bancarias (tipo, plazo, carencia)",
            "Firmar contratos de financiación",
        ],
        "3. Licencias": [
            "Licencia de actividad / apertura",
            "Licencia de obras (si reformas)",
            "Proyecto técnico (arquitecto/ingeniero)",
            "Certificado de solidez / estructura",
            "Autorización sanitaria (Registro Sanitario)",
            "Inscripción RGSEAA (si elaboración)",
            "Plan de autoprotección (si aforo > 50)",
            "Contrato de gestión de residuos",
        ],
        "4. Proveedores": [
            "Seleccionar distribuidor principal (alimentación)",
            "Contratar suministros (luz, agua, gas)",
            "Negociar contrato de alquiler y fianza",
            "Contratar servicio de lavandería",
            "Contratar mantenimiento de equipos",
            "Contratar servicio de desratización (DDD)",
            "Firmar contrato con plataformas delivery",
            "Negociar condiciones de pago con proveedores clave",
        ],
        "5. Seguros": [
            "Seguro de responsabilidad civil",
            "Seguro multirriesgo del local",
            "Seguro de accidentes de trabajo",
            "Seguro de pérdida de beneficios",
            "Seguro de mercancías (cámaras frías)",
            "Seguro de responsabilidad de administradores",
            "Seguro de ciberriesgos (si TPV/web)",
            "Seguro de caución (si aplica)",
        ],
        "6. Tesorería": [
            "Configurar TPV y datáfonos",
            "Activar pasarela de pago online",
            "Configurar software de facturación (Veri*factu)",
            "Establecer fondo de caja (efectivo inicial)",
            "Definir política de propinas",
            "Configurar domiciliaciones (alquiler, suministros)",
            "Crear previsión de tesorería mes 1-3",
            "Establecer umbral de alerta de tesorería",
        ],
    }

    # Phase colors
    phase_colors = {
        "1. Constitución": "E3F2FD",
        "2. Financiación": "FFF3E0",
        "3. Licencias": "E8F5E9",
        "4. Proveedores": "F3E5F5",
        "5. Seguros": "FFEBEE",
        "6. Tesorería": "E0F7FA",
    }

    r = 5
    task_num = 1
    for phase, items in tasks.items():
        phase_fill = PatternFill(start_color=phase_colors[phase], end_color=phase_colors[phase], fill_type="solid")
        for item in items:
            ws.cell(row=r, column=1, value=task_num).font = data_font
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=1).alignment = center_align

            ws.cell(row=r, column=2, value=phase).font = data_font
            ws.cell(row=r, column=2).border = thin_border
            ws.cell(row=r, column=2).fill = phase_fill

            ws.cell(row=r, column=3, value=item).font = data_font
            ws.cell(row=r, column=3).border = thin_border

            ws.cell(row=r, column=4).border = thin_border
            ws.cell(row=r, column=4).fill = input_fill
            ws.cell(row=r, column=4).font = data_font

            ws.cell(row=r, column=5).border = thin_border
            ws.cell(row=r, column=5).fill = input_fill
            ws.cell(row=r, column=5).number_format = 'DD/MM/YYYY'
            ws.cell(row=r, column=5).font = data_font

            ws.cell(row=r, column=6, value="Pendiente").font = data_font
            ws.cell(row=r, column=6).border = thin_border
            ws.cell(row=r, column=6).alignment = center_align
            dv.add(ws.cell(row=r, column=6))

            ws.cell(row=r, column=7).border = thin_border
            ws.cell(row=r, column=7).fill = input_fill
            ws.cell(row=r, column=7).font = data_font

            task_num += 1
            r += 1

    # Summary
    r += 1
    ws.cell(row=r, column=2, value="RESUMEN").font = section_font
    r += 1
    ws.cell(row=r, column=2, value="Total tareas:").font = bold_font
    ws.cell(row=r, column=3, value=48).font = bold_font
    r += 1
    ws.cell(row=r, column=2, value="Completadas:").font = bold_font
    ws.cell(row=r, column=3).font = formula_font
    ws.cell(row=r, column=3).value = '=COUNTIF(F5:F52,"Completada")'
    r += 1
    ws.cell(row=r, column=2, value="En curso:").font = bold_font
    ws.cell(row=r, column=3).font = formula_font
    ws.cell(row=r, column=3).value = '=COUNTIF(F5:F52,"En curso")'
    r += 1
    ws.cell(row=r, column=2, value="Pendientes:").font = bold_font
    ws.cell(row=r, column=3).font = formula_font
    ws.cell(row=r, column=3).value = '=COUNTIF(F5:F52,"Pendiente")'
    r += 1
    ws.cell(row=r, column=2, value="% Completado:").font = bold_font
    ws.cell(row=r, column=3).font = Font(name="Calibri", size=13, bold=True, color=GOLD)
    ws.cell(row=r, column=3).value = '=COUNTIF(F5:F52,"Completada")/48'
    ws.cell(row=r, column=3).number_format = PCT_FMT

    add_brand_footer(ws, r + 2, "G")
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-09-checklist-pre-apertura.xlsx"))
    print("  ✓ BONUS-09-checklist-pre-apertura.xlsx")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\n📁 Generando Kit Plan Financiero en: {OUTPUT_DIR}\n")
    gen_01()
    gen_01b()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_08()
    gen_09()
    print(f"\n✅ 10 archivos generados correctamente en {OUTPUT_DIR}\n")
