#!/usr/bin/env python3
"""Generate 9 Excel templates for Kit Control de Inventario y Compras."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(__file__), '..', 'public', 'dl', 'kit-inventario')
os.makedirs(OUT, exist_ok=True)

# ── Brand colors ──
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
RED = "FF4444"
AMBER = "FFB800"
GREEN = "00C853"

ZONE_COLORS = {
    'Cárnicos': 'E53935',
    'Pescados': '1E88E5',
    'Lácteos': 'FDD835',
    'Verduras': '43A047',
    'Frutas': 'FB8C00',
    'Secos/Granos': '8D6E63',
    'Congelados': '29B6F6',
    'Bebidas': '7E57C2',
    'Limpieza': '78909C',
    'Varios': '90A4AE',
}

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY),
)

def style_header(ws, row, cols, bg=HEADER_BG):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data_row(ws, row, cols, fill_color=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(size=10)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)

def add_instructions(wb, product_name, content_lines):
    ws = wb.create_sheet("Instrucciones", 0)
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions['A'].width = 80
    ws.cell(1, 1, product_name).font = Font(bold=True, size=16, color=GOLD)
    ws.cell(2, 1, "AI Chef Pro — aichef.pro").font = Font(size=10, color="888888")
    ws.cell(3, 1, "").font = Font(size=10)
    for i, line in enumerate(content_lines, 4):
        ws.cell(i, 1, line).font = Font(size=10)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)

def add_footer(ws, row, cols):
    ws.cell(row, 1, "© 2026 AI Chef Pro · aichef.pro").font = Font(size=8, color="999999")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(cols, 8))

def add_check_validation(ws, col_letter, start_row, end_row):
    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Solo ✓, ✗ o —"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


# ═══════════════════════════════════════════════════════════════
# 01 — Inventario de Stock Diario
# ═══════════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Inventario de Stock Diario", [
        "Este archivo te permite controlar el stock diario de tu restaurante.",
        "Cada pestaña corresponde a una zona: Cocina, Barra, Almacén.",
        "",
        "CÓMO USAR:",
        "1. Rellena la columna 'Producto' con tus ingredientes.",
        "2. Establece 'Par Level' (stock mínimo) y 'Par Max' (stock máximo).",
        "3. Cada día, anota el stock actual en 'Stock Real'.",
        "4. La columna 'Estado' se colorea automáticamente:",
        "   🔴 Rojo = por debajo del par level → ¡pedir!",
        "   🟡 Amarillo = stock bajo, vigilar",
        "   🟢 Verde = stock correcto",
        "5. 'A Pedir' calcula automáticamente cuánto necesitas.",
        "",
        "PERSONALIZACIÓN:",
        "- Añade o elimina filas según tus productos.",
        "- Ajusta los par levels a tu ritmo de ventas.",
        "- La pestaña 'Resumen' totaliza automáticamente.",
    ])

    zones = [
        ('Cocina', ['Pechuga de pollo', 'Solomillo ternera', 'Salmón fresco', 'Gambas', 'Tomate', 'Cebolla', 'Lechuga', 'Patata', 'Aceite oliva', 'Sal', 'Pimienta', 'Arroz', 'Pasta seca', 'Harina', 'Huevos', 'Nata', 'Mantequilla', 'Queso parmesano', 'Limón', 'Ajo']),
        ('Barra', ['Café en grano', 'Leche entera', 'Leche avena', 'Zumo naranja', 'Coca-Cola', 'Cerveza grifo (barril)', 'Agua mineral', 'Tónica', 'Vino tinto (copa)', 'Vino blanco (copa)', 'Hielo', 'Azúcar', 'Servilletas', 'Pajitas papel', 'Vasos desechables']),
        ('Almacén', ['Papel cocina', 'Film transparente', 'Papel aluminio', 'Bolsas basura', 'Guantes nitrilo', 'Lavavajillas', 'Desengrasante', 'Lejía alimentaria', 'Contenedores GN 1/1', 'Etiquetas FIFO', 'Rollos de ticket', 'Servilletas extra', 'Aceite girasol (5L)', 'Vinagre', 'Legumbres secas']),
    ]

    for zone_name, products in zones:
        ws = wb.create_sheet(zone_name)
        ws.sheet_properties.tabColor = GOLD
        headers = ['#', 'Producto', 'Categoría', 'Unidad', 'Par Level', 'Par Max', 'Stock Real', 'Estado', 'A Pedir', 'Valor (€)', 'Proveedor', 'Notas']
        widths = [5, 25, 14, 8, 10, 10, 10, 10, 10, 10, 18, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.cell(1, 1, f"INVENTARIO — {zone_name.upper()}").font = Font(bold=True, size=14, color=GOLD)
        ws.merge_cells('A1:L1')
        ws.cell(2, 1, "Fecha: ___/___/______    Responsable: ________________").font = Font(size=10, color="666666")
        ws.merge_cells('A2:L2')

        row = 4
        for i, h in enumerate(headers, 1):
            ws.cell(row, i, h)
        style_header(ws, row, len(headers))

        categories = list(ZONE_COLORS.keys())
        for idx, product in enumerate(products, 1):
            r = row + idx
            cat = categories[idx % len(categories)]
            ws.cell(r, 1, idx)
            ws.cell(r, 2, product)
            ws.cell(r, 3, cat)
            ws.cell(r, 4, 'kg' if idx % 3 == 0 else 'ud' if idx % 3 == 1 else 'L')
            ws.cell(r, 5, 5)  # par level
            ws.cell(r, 6, 15)  # par max
            ws.cell(r, 7, '')  # stock real
            ws.cell(r, 8).value = f'=IF(G{r}="","",IF(G{r}<E{r},"🔴 PEDIR",IF(G{r}<E{r}*1.5,"🟡 BAJO","🟢 OK")))'
            ws.cell(r, 9).value = f'=IF(G{r}="","",IF(G{r}<E{r},F{r}-G{r},0))'
            style_data_row(ws, r, len(headers), LIGHT_GRAY if idx % 2 == 0 else None)

        add_footer(ws, row + len(products) + 3, len(headers))

    # Resumen
    ws = wb.create_sheet("Resumen Dashboard")
    ws.sheet_properties.tabColor = GREEN
    ws.cell(1, 1, "RESUMEN DE STOCK").font = Font(bold=True, size=14, color=GOLD)
    ws.cell(3, 1, "Zona").font = Font(bold=True, size=11)
    ws.cell(3, 2, "Productos con stock bajo").font = Font(bold=True, size=11)
    ws.cell(3, 3, "Productos a pedir").font = Font(bold=True, size=11)
    ws.cell(3, 4, "Valor total stock (€)").font = Font(bold=True, size=11)
    for i, w in enumerate([20, 25, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(os.path.join(OUT, '01-inventario-stock-diario.xlsx'))
    print("✅ 01-inventario-stock-diario.xlsx")


# ═══════════════════════════════════════════════════════════════
# 02 — Fichas de Proveedores
# ═══════════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Fichas de Proveedores", [
        "Gestiona todos tus proveedores y compara precios en un solo archivo.",
        "",
        "PESTAÑAS:",
        "1. Directorio — datos de contacto de todos tus proveedores",
        "2. Comparativa Precios — compara precios de 5 proveedores por producto",
        "3. Evaluación — puntúa calidad, precio, entrega, servicio (1-5)",
        "4. Condiciones Comerciales — plazos de pago, mínimos, rappels",
    ])

    # Directorio
    ws = wb.create_sheet("Directorio Proveedores")
    ws.sheet_properties.tabColor = GOLD
    headers = ['#', 'Proveedor', 'Categoría', 'Contacto', 'Teléfono', 'Email', 'Dirección', 'Días Entrega', 'Pedido Mínimo (€)', 'Forma Pago', 'Notas']
    for i, w in enumerate([5, 25, 14, 18, 14, 22, 25, 14, 14, 14, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "DIRECTORIO DE PROVEEDORES").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:K1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    for r in range(4, 24):
        ws.cell(r, 1, r - 3)
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    # Comparativa
    ws = wb.create_sheet("Comparativa Precios")
    ws.sheet_properties.tabColor = AMBER
    headers = ['#', 'Producto', 'Unidad', 'Prov. 1', 'Prov. 2', 'Prov. 3', 'Prov. 4', 'Prov. 5', 'Mejor Precio', 'Mejor Proveedor', '% Diferencia']
    for i, w in enumerate([5, 25, 8, 12, 12, 12, 12, 12, 12, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "COMPARATIVA DE PRECIOS POR PROVEEDOR").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:K1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    products = ['Pechuga pollo (kg)', 'Solomillo ternera (kg)', 'Salmón fresco (kg)', 'Aceite oliva AOVE (5L)', 'Tomate (kg)', 'Lechuga (ud)', 'Patata (kg)', 'Huevos M (docena)', 'Leche entera (L)', 'Café grano (kg)']
    for idx, prod in enumerate(products, 1):
        r = idx + 3
        ws.cell(r, 1, idx)
        ws.cell(r, 2, prod)
        ws.cell(r, 3, 'kg' if 'kg' in prod else 'ud' if 'ud' in prod else 'L' if 'L' in prod else 'ud')
        ws.cell(r, 9).value = f'=MIN(D{r}:H{r})'
        ws.cell(r, 11).value = f'=IF(I{r}=0,"",ROUND((MAX(D{r}:H{r})-I{r})/I{r}*100,1))'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if idx % 2 == 0 else None)

    # Evaluación
    ws = wb.create_sheet("Evaluación Proveedores")
    ws.sheet_properties.tabColor = GREEN
    headers = ['#', 'Proveedor', 'Calidad (1-5)', 'Precio (1-5)', 'Puntualidad (1-5)', 'Servicio (1-5)', 'Flexibilidad (1-5)', 'TOTAL', 'Nota']
    for i, w in enumerate([5, 25, 14, 14, 14, 14, 14, 10, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "EVALUACIÓN DE PROVEEDORES").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:I1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    for r in range(4, 14):
        ws.cell(r, 1, r - 3)
        ws.cell(r, 8).value = f'=IF(C{r}="","",AVERAGE(C{r}:G{r}))'
        ws.cell(r, 8).number_format = '0.0'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    # Condiciones
    ws = wb.create_sheet("Condiciones Comerciales")
    ws.sheet_properties.tabColor = "7E57C2"
    headers = ['#', 'Proveedor', 'Plazo Pago', 'Pedido Mínimo', 'Rappel', 'Transporte', 'Horario Entrega', 'Días Pedido', 'Días Entrega', 'Notas']
    for i, w in enumerate([5, 25, 12, 14, 14, 14, 14, 14, 14, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "CONDICIONES COMERCIALES").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:J1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    for r in range(4, 14):
        ws.cell(r, 1, r - 3)
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    wb.save(os.path.join(OUT, '02-fichas-proveedores.xlsx'))
    print("✅ 02-fichas-proveedores.xlsx")


# ═══════════════════════════════════════════════════════════════
# 03 — Pedidos de Compra
# ═══════════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Pedidos de Compra", [
        "Genera pedidos de compra profesionales para tus proveedores.",
        "",
        "1. Rellena el proveedor y los productos en la pestaña 'Pedido Actual'.",
        "2. Los subtotales e IVA se calculan automáticamente.",
        "3. Usa la pestaña 'Historial' para llevar registro de todos los pedidos.",
        "4. La pestaña 'Imprimible' está optimizada para enviar o imprimir.",
    ])

    ws = wb.create_sheet("Pedido Actual")
    ws.sheet_properties.tabColor = GOLD
    ws.cell(1, 1, "PEDIDO DE COMPRA").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:H1')
    ws.cell(3, 1, "Proveedor:").font = Font(bold=True, size=11)
    ws.cell(3, 3, "").font = Font(size=11)
    ws.cell(4, 1, "Fecha pedido:").font = Font(bold=True, size=11)
    ws.cell(4, 3, "").font = Font(size=11)
    ws.cell(5, 1, "Fecha entrega:").font = Font(bold=True, size=11)
    ws.cell(5, 3, "").font = Font(size=11)
    ws.cell(6, 1, "Nº Pedido:").font = Font(bold=True, size=11)

    headers = ['#', 'Producto', 'Unidad', 'Cantidad', 'Precio/ud (€)', 'Subtotal (€)', 'IVA %', 'Total (€)']
    for i, w in enumerate([5, 30, 8, 10, 14, 14, 8, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers, 1):
        ws.cell(8, i, h)
    style_header(ws, 8, len(headers))

    for r in range(9, 29):
        ws.cell(r, 1, r - 8)
        ws.cell(r, 6).value = f'=IF(D{r}="","",D{r}*E{r})'
        ws.cell(r, 6).number_format = '#,##0.00'
        ws.cell(r, 7, 10)
        ws.cell(r, 8).value = f'=IF(F{r}="","",F{r}*(1+G{r}/100))'
        ws.cell(r, 8).number_format = '#,##0.00'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 8) % 2 == 0 else None)

    ws.cell(30, 5, "SUBTOTAL:").font = Font(bold=True, size=11)
    ws.cell(30, 6).value = '=SUM(F9:F28)'
    ws.cell(30, 6).number_format = '#,##0.00'
    ws.cell(30, 6).font = Font(bold=True, size=11)
    ws.cell(31, 5, "TOTAL CON IVA:").font = Font(bold=True, size=12, color=GOLD)
    ws.cell(31, 6).value = '=SUM(H9:H28)'
    ws.cell(31, 6).number_format = '#,##0.00'
    ws.cell(31, 6).font = Font(bold=True, size=12, color=GOLD)

    # Historial
    ws2 = wb.create_sheet("Historial Pedidos")
    ws2.sheet_properties.tabColor = "78909C"
    headers2 = ['Fecha', 'Nº Pedido', 'Proveedor', 'Productos', 'Subtotal (€)', 'IVA (€)', 'Total (€)', 'Estado', 'Observaciones']
    for i, w in enumerate([12, 12, 20, 30, 12, 10, 12, 12, 25], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.cell(1, 1, "HISTORIAL DE PEDIDOS").font = Font(bold=True, size=14, color=GOLD)
    ws2.merge_cells('A1:I1')
    for i, h in enumerate(headers2, 1):
        ws2.cell(3, i, h)
    style_header(ws2, 3, len(headers2))

    wb.save(os.path.join(OUT, '03-pedidos-compra.xlsx'))
    print("✅ 03-pedidos-compra.xlsx")


# ═══════════════════════════════════════════════════════════════
# 04 — Recepción de Mercancías
# ═══════════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Recepción de Mercancías", [
        "Verifica cada entrega de proveedor con este checklist APPCC.",
        "",
        "1. Cada vez que llega un proveedor, abre la pestaña 'Control Recepción'.",
        "2. Verifica: cantidad, calidad visual, temperatura y fecha caducidad.",
        "3. Registra incidencias en la pestaña correspondiente.",
        "4. Las temperaturas fuera de rango se marcan automáticamente en rojo.",
    ])

    ws = wb.create_sheet("Control Recepción")
    ws.sheet_properties.tabColor = GOLD
    headers = ['Fecha', 'Proveedor', 'Producto', 'Pedido (ud)', 'Recibido (ud)', '✓ Cantidad', 'Temp. °C', '✓ Temp. OK', 'Caducidad', '✓ Calidad', 'Incidencia', 'Firmado']
    for i, w in enumerate([12, 20, 22, 10, 10, 10, 10, 10, 12, 10, 20, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "CONTROL DE RECEPCIÓN DE MERCANCÍAS").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:L1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    add_check_validation(ws, 'F', 4, 33)
    add_check_validation(ws, 'H', 4, 33)
    add_check_validation(ws, 'J', 4, 33)
    for r in range(4, 34):
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    # Incidencias
    ws2 = wb.create_sheet("Registro Incidencias")
    ws2.sheet_properties.tabColor = RED
    headers2 = ['Fecha', 'Proveedor', 'Producto', 'Tipo Incidencia', 'Descripción', 'Acción Tomada', 'Responsable', 'Resuelta']
    for i, w in enumerate([12, 20, 22, 18, 30, 25, 14, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.cell(1, 1, "REGISTRO DE INCIDENCIAS — RECEPCIÓN").font = Font(bold=True, size=14, color=RED)
    ws2.merge_cells('A1:H1')
    for i, h in enumerate(headers2, 1):
        ws2.cell(3, i, h)
    style_header(ws2, 3, len(headers2), RED)

    # Temperaturas
    ws3 = wb.create_sheet("Verificación Temperaturas")
    ws3.sheet_properties.tabColor = "1E88E5"
    ws3.cell(1, 1, "TEMPERATURAS DE REFERENCIA — RECEPCIÓN").font = Font(bold=True, size=14, color=GOLD)
    ws3.merge_cells('A1:D1')
    refs = [
        ('Carne fresca', '0 a 4°C', '7°C máx.'),
        ('Pescado fresco', '0 a 2°C', '4°C máx.'),
        ('Lácteos', '0 a 6°C', '8°C máx.'),
        ('Congelados', '-18°C o menos', '-15°C máx.'),
        ('Frutas y verduras', '4 a 8°C', '12°C máx.'),
        ('Huevos', '< 18°C transporte', '20°C máx.'),
    ]
    headers3 = ['Categoría', 'Temp. Ideal', 'Temp. Máx. Aceptable']
    for i, h in enumerate(headers3, 1):
        ws3.cell(3, i, h)
    style_header(ws3, 3, 3)
    for i, (cat, ideal, max_t) in enumerate(refs, 1):
        r = i + 3
        ws3.cell(r, 1, cat)
        ws3.cell(r, 2, ideal)
        ws3.cell(r, 3, max_t)
        style_data_row(ws3, r, 3, LIGHT_GRAY if i % 2 == 0 else None)
    for c in range(1, 4):
        ws3.column_dimensions[get_column_letter(c)].width = 22

    wb.save(os.path.join(OUT, '04-recepcion-mercancias.xlsx'))
    print("✅ 04-recepcion-mercancias.xlsx")


# ═══════════════════════════════════════════════════════════════
# 05 — Control de Mermas
# ═══════════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Control de Mermas", [
        "Registra y analiza todas las mermas de tu restaurante.",
        "",
        "1. Anota cada merma en 'Registro Diario': producto, cantidad, motivo.",
        "2. El coste se calcula automáticamente.",
        "3. 'Análisis por Categoría' muestra dónde pierdes más dinero.",
        "4. 'Dashboard' resume las métricas clave del mes.",
        "5. 'Plan de Acción' te ayuda a reducir mermas sistemáticamente.",
        "",
        "OBJETIVO: mantener mermas por debajo del 3% sobre compras.",
    ])

    ws = wb.create_sheet("Registro Diario Mermas")
    ws.sheet_properties.tabColor = RED
    headers = ['Fecha', 'Producto', 'Categoría', 'Cantidad', 'Unidad', 'Coste/ud (€)', 'Coste Total (€)', 'Motivo', 'Responsable', 'Acción Correctiva']
    for i, w in enumerate([12, 22, 14, 10, 8, 12, 12, 18, 14, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "REGISTRO DIARIO DE MERMAS").font = Font(bold=True, size=14, color=RED)
    ws.merge_cells('A1:J1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers), RED)

    motivos_dv = DataValidation(type="list", formula1='"Caducidad,Preparación,Devolución,Accidente,Sobreproducción,Deterioro,Otro"')
    ws.add_data_validation(motivos_dv)

    for r in range(4, 54):
        ws.cell(r, 7).value = f'=IF(D{r}="","",D{r}*F{r})'
        ws.cell(r, 7).number_format = '#,##0.00'
        motivos_dv.add(f"H{r}")
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    ws.cell(55, 6, "TOTAL MES:").font = Font(bold=True, size=12)
    ws.cell(55, 7).value = '=SUM(G4:G53)'
    ws.cell(55, 7).number_format = '#,##0.00'
    ws.cell(55, 7).font = Font(bold=True, size=12, color=RED)

    # Análisis
    ws2 = wb.create_sheet("Análisis por Categoría")
    ws2.sheet_properties.tabColor = AMBER
    headers2 = ['Categoría', 'Nº Incidencias', 'Coste Total (€)', '% del Total', 'Tendencia vs Mes Anterior']
    for i, w in enumerate([18, 16, 16, 14, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.cell(1, 1, "ANÁLISIS DE MERMAS POR CATEGORÍA").font = Font(bold=True, size=14, color=GOLD)
    ws2.merge_cells('A1:E1')
    for i, h in enumerate(headers2, 1):
        ws2.cell(3, i, h)
    style_header(ws2, 3, len(headers2))
    for i, cat in enumerate(['Cárnicos', 'Pescados', 'Lácteos', 'Verduras/Frutas', 'Secos', 'Congelados', 'Bebidas'], 1):
        ws2.cell(i + 3, 1, cat)
        style_data_row(ws2, i + 3, len(headers2), LIGHT_GRAY if i % 2 == 0 else None)

    # Dashboard
    ws3 = wb.create_sheet("Dashboard Mermas")
    ws3.sheet_properties.tabColor = GREEN
    ws3.cell(1, 1, "DASHBOARD DE MERMAS — MES").font = Font(bold=True, size=14, color=GOLD)
    ws3.cell(3, 1, "KPI").font = Font(bold=True)
    ws3.cell(3, 2, "Valor").font = Font(bold=True)
    ws3.cell(3, 3, "Objetivo").font = Font(bold=True)
    ws3.cell(3, 4, "Estado").font = Font(bold=True)
    kpis = [
        ('Coste total mermas (€)', '', '< 500 €', ''),
        ('% mermas / compras', '', '< 3%', ''),
        ('Nº incidencias', '', '< 20', ''),
        ('Categoría con más merma', '', '-', ''),
        ('Motivo más frecuente', '', '-', ''),
    ]
    for i, (kpi, val, obj, estado) in enumerate(kpis, 1):
        r = i + 3
        ws3.cell(r, 1, kpi)
        ws3.cell(r, 2, val)
        ws3.cell(r, 3, obj)
        ws3.cell(r, 4, estado)
    for c in range(1, 5):
        ws3.column_dimensions[get_column_letter(c)].width = 22

    # Plan de Acción
    ws4 = wb.create_sheet("Plan de Acción")
    ws4.sheet_properties.tabColor = "43A047"
    headers4 = ['Prioridad', 'Categoría', 'Problema Identificado', 'Acción Correctiva', 'Responsable', 'Fecha Límite', 'Estado']
    for i, w in enumerate([10, 14, 25, 30, 14, 14, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.cell(1, 1, "PLAN DE ACCIÓN — REDUCCIÓN DE MERMAS").font = Font(bold=True, size=14, color=GOLD)
    ws4.merge_cells('A1:G1')
    for i, h in enumerate(headers4, 1):
        ws4.cell(3, i, h)
    style_header(ws4, 3, len(headers4))

    wb.save(os.path.join(OUT, '05-control-mermas.xlsx'))
    print("✅ 05-control-mermas.xlsx")


# ═══════════════════════════════════════════════════════════════
# 06 — FIFO y Caducidades
# ═══════════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "FIFO y Caducidades", [
        "Controla la rotación First In First Out y las caducidades.",
        "",
        "1. Registra cada producto con su fecha de caducidad.",
        "2. El semáforo de colores te avisa automáticamente:",
        "   🔴 Rojo = caduca mañana o ya caducado",
        "   🟡 Amarillo = caduca en 3 días",
        "   🟢 Verde = OK",
        "3. Usa el 'Mapa de Almacén' para organizar por zonas.",
    ])

    ws = wb.create_sheet("Control FIFO")
    ws.sheet_properties.tabColor = GOLD
    headers = ['#', 'Producto', 'Categoría', 'Lote/Ref', 'Fecha Entrada', 'Fecha Caducidad', 'Días Restantes', 'Estado', 'Zona Almacén', 'Posición', 'Acción']
    for i, w in enumerate([5, 22, 14, 12, 14, 14, 14, 14, 14, 10, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "CONTROL FIFO — FIRST IN FIRST OUT").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:K1')
    ws.cell(2, 1, "Fecha revisión: ___/___/______    Responsable: ________________").font = Font(size=10, color="666666")
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    for r in range(5, 55):
        ws.cell(r, 1, r - 4)
        ws.cell(r, 7).value = f'=IF(F{r}="","",F{r}-TODAY())'
        ws.cell(r, 8).value = f'=IF(G{r}="","",IF(G{r}<=1,"🔴 URGENTE",IF(G{r}<=3,"🟡 PRONTO","🟢 OK")))'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 4) % 2 == 0 else None)

    # Alertas
    ws2 = wb.create_sheet("Alertas Caducidad")
    ws2.sheet_properties.tabColor = RED
    ws2.cell(1, 1, "⚠️ ALERTAS DE CADUCIDAD").font = Font(bold=True, size=14, color=RED)
    ws2.cell(3, 1, "Filtra la pestaña 'Control FIFO' por Estado = 🔴 URGENTE o 🟡 PRONTO").font = Font(size=11)
    ws2.cell(5, 1, "PROTOCOLO DE ACTUACIÓN:").font = Font(bold=True, size=12, color=GOLD)
    protocols = [
        "1. Productos con 🔴 URGENTE → usar HOY o descartar y registrar en Control de Mermas.",
        "2. Productos con 🟡 PRONTO → priorizar en la producción de los próximos 3 días.",
        "3. Revisar esta pestaña al inicio de cada turno de mañana.",
        "4. Anotar en el briefing de servicio los productos a consumir con prioridad.",
    ]
    for i, p in enumerate(protocols, 7):
        ws2.cell(i, 1, p).font = Font(size=10)
    ws2.column_dimensions['A'].width = 70

    # Mapa Almacén
    ws3 = wb.create_sheet("Mapa Almacén")
    ws3.sheet_properties.tabColor = "43A047"
    ws3.cell(1, 1, "MAPA DE ALMACÉN — ZONAS DE ALMACENAMIENTO").font = Font(bold=True, size=14, color=GOLD)
    headers3 = ['Zona', 'Temperatura', 'Categorías', 'Capacidad', 'Notas']
    for i, w in enumerate([16, 14, 30, 14, 25], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers3, 1):
        ws3.cell(3, i, h)
    style_header(ws3, 3, len(headers3))
    zones = [
        ('Cámara Fría 1', '0-4°C', 'Cárnicos, Lácteos', '', ''),
        ('Cámara Fría 2', '0-2°C', 'Pescados, Mariscos', '', ''),
        ('Congelador', '-18°C', 'Congelados', '', ''),
        ('Almacén Seco', '15-20°C', 'Secos, Conservas, Bebidas', '', ''),
        ('Zona Frutas', '4-8°C', 'Verduras, Frutas', '', ''),
        ('Bodega', '12-16°C', 'Vinos', '', ''),
    ]
    for i, (zona, temp, cats, cap, notas) in enumerate(zones, 1):
        r = i + 3
        ws3.cell(r, 1, zona)
        ws3.cell(r, 2, temp)
        ws3.cell(r, 3, cats)
        style_data_row(ws3, r, len(headers3), LIGHT_GRAY if i % 2 == 0 else None)

    wb.save(os.path.join(OUT, '06-fifo-caducidades.xlsx'))
    print("✅ 06-fifo-caducidades.xlsx")


# ═══════════════════════════════════════════════════════════════
# 07 — Análisis de Costes de Compras
# ═══════════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Análisis de Costes de Compras", [
        "Analiza y controla los costes de compras de tu restaurante.",
        "",
        "PESTAÑAS:",
        "1. Coste por Categoría — desglose mensual por tipo de producto",
        "2. Evolución Mensual — tendencia de costes 12 meses",
        "3. Top 20 Productos — los productos que más dinero cuestan",
        "4. Dashboard KPIs — métricas clave de compras",
    ])

    # Coste por Categoría
    ws = wb.create_sheet("Coste por Categoría")
    ws.sheet_properties.tabColor = GOLD
    headers = ['Categoría', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic', 'Total Año', '% del Total']
    for i, w in enumerate([18] + [10]*12 + [12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "COSTE DE COMPRAS POR CATEGORÍA").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:O1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    categories = ['Cárnicos', 'Pescados', 'Lácteos', 'Verduras/Frutas', 'Secos/Granos', 'Congelados', 'Bebidas Alcohólicas', 'Bebidas No Alcohólicas', 'Limpieza', 'Otros']
    for i, cat in enumerate(categories, 1):
        r = i + 3
        ws.cell(r, 1, cat)
        ws.cell(r, 14).value = f'=SUM(B{r}:M{r})'
        ws.cell(r, 14).number_format = '#,##0.00'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if i % 2 == 0 else None)
    total_r = len(categories) + 4
    ws.cell(total_r, 1, "TOTAL").font = Font(bold=True, size=11)
    for c in range(2, 15):
        ws.cell(total_r, c).value = f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}{total_r-1})'
        ws.cell(total_r, c).number_format = '#,##0.00'
        ws.cell(total_r, c).font = Font(bold=True)

    # Top 20
    ws2 = wb.create_sheet("Top 20 Productos")
    ws2.sheet_properties.tabColor = AMBER
    headers2 = ['#', 'Producto', 'Categoría', 'Gasto Mensual (€)', 'Gasto Anual (€)', '% del Total', 'Precio Actual', 'Variación vs Anterior', 'Proveedor']
    for i, w in enumerate([5, 25, 14, 16, 14, 10, 14, 16, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.cell(1, 1, "TOP 20 PRODUCTOS — MAYOR GASTO").font = Font(bold=True, size=14, color=GOLD)
    ws2.merge_cells('A1:I1')
    for i, h in enumerate(headers2, 1):
        ws2.cell(3, i, h)
    style_header(ws2, 3, len(headers2))
    for r in range(4, 24):
        ws2.cell(r, 1, r - 3)
        style_data_row(ws2, r, len(headers2), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    # Dashboard
    ws3 = wb.create_sheet("Dashboard KPIs")
    ws3.sheet_properties.tabColor = GREEN
    ws3.cell(1, 1, "DASHBOARD — KPIs DE COMPRAS").font = Font(bold=True, size=14, color=GOLD)
    kpis = [
        ('Food Cost %', '', '28-32%', 'Coste materia prima / Ventas × 100'),
        ('Coste por Cubierto (€)', '', '< 4,50 €', 'Compras del mes / Nº cubiertos'),
        ('Gasto Total Compras (€)', '', '-', 'Suma todas las categorías'),
        ('Variación vs Mes Anterior', '', '< ±5%', 'Diferencia porcentual'),
        ('Nº Proveedores Activos', '', '-', 'Proveedores con pedidos este mes'),
        ('Producto con Mayor Variación', '', '-', 'Alerta si sube >5%'),
    ]
    ws3.cell(3, 1, "KPI").font = Font(bold=True, size=11)
    ws3.cell(3, 2, "Valor Actual").font = Font(bold=True, size=11)
    ws3.cell(3, 3, "Objetivo").font = Font(bold=True, size=11)
    ws3.cell(3, 4, "Cómo Calcular").font = Font(bold=True, size=11)
    style_header(ws3, 3, 4)
    for i, (kpi, val, obj, calc) in enumerate(kpis, 1):
        r = i + 3
        ws3.cell(r, 1, kpi)
        ws3.cell(r, 2, val)
        ws3.cell(r, 3, obj)
        ws3.cell(r, 4, calc)
        style_data_row(ws3, r, 4, LIGHT_GRAY if i % 2 == 0 else None)
    for c in range(1, 5):
        ws3.column_dimensions[get_column_letter(c)].width = 25

    wb.save(os.path.join(OUT, '07-analisis-costes-compras.xlsx'))
    print("✅ 07-analisis-costes-compras.xlsx")


# ═══════════════════════════════════════════════════════════════
# BONUS 08 — Inventario Rápido Mensual
# ═══════════════════════════════════════════════════════════════
def gen_bonus_08():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "BONUS: Inventario Rápido Mensual", [
        "Hoja simplificada para hacer el inventario mensual completo.",
        "",
        "1. Imprime esta hoja y recorre tu almacén contando productos.",
        "2. Anota cantidad y la hoja calcula el valor automáticamente.",
        "3. Compara con el mes anterior para detectar variaciones.",
    ])

    ws = wb.create_sheet("Conteo Rápido")
    ws.sheet_properties.tabColor = GOLD
    headers = ['#', 'Producto', 'Categoría', 'Unidad', 'Precio/ud (€)', 'Stock Actual', 'Valor (€)', 'Stock Mes Anterior', 'Variación']
    for i, w in enumerate([5, 25, 14, 8, 14, 12, 14, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "INVENTARIO RÁPIDO MENSUAL").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:I1')
    ws.cell(2, 1, "Mes: ________________    Fecha: ___/___/______    Responsable: ________________").font = Font(size=10, color="666666")
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    for r in range(5, 55):
        ws.cell(r, 1, r - 4)
        ws.cell(r, 7).value = f'=IF(F{r}="","",E{r}*F{r})'
        ws.cell(r, 7).number_format = '#,##0.00'
        ws.cell(r, 9).value = f'=IF(F{r}="","",F{r}-H{r})'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 4) % 2 == 0 else None)
    ws.cell(56, 6, "TOTAL:").font = Font(bold=True, size=12)
    ws.cell(56, 7).value = '=SUM(G5:G54)'
    ws.cell(56, 7).number_format = '#,##0.00'
    ws.cell(56, 7).font = Font(bold=True, size=12, color=GOLD)

    wb.save(os.path.join(OUT, 'BONUS-08-inventario-rapido-mensual.xlsx'))
    print("✅ BONUS-08-inventario-rapido-mensual.xlsx")


# ═══════════════════════════════════════════════════════════════
# BONUS 09 — Calculadora Punto de Pedido
# ═══════════════════════════════════════════════════════════════
def gen_bonus_09():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "BONUS: Calculadora Punto de Pedido", [
        "Calcula el punto óptimo de reposición para cada producto.",
        "",
        "FÓRMULA:",
        "Punto de Pedido = (Consumo Diario × Lead Time) + Stock Seguridad",
        "",
        "1. Introduce consumo diario, tiempo de entrega del proveedor y stock de seguridad.",
        "2. La calculadora te dice exactamente cuándo pedir.",
        "3. El simulador te permite probar diferentes escenarios de demanda.",
    ])

    ws = wb.create_sheet("Calculadora")
    ws.sheet_properties.tabColor = GOLD
    headers = ['#', 'Producto', 'Consumo Diario', 'Lead Time (días)', 'Stock Seguridad', 'Punto de Pedido', 'EOQ Sugerida', 'Frecuencia Pedido (días)', 'Proveedor']
    for i, w in enumerate([5, 25, 14, 16, 14, 14, 14, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(1, 1, "CALCULADORA — PUNTO DE PEDIDO ÓPTIMO").font = Font(bold=True, size=14, color=GOLD)
    ws.merge_cells('A1:I1')
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    for r in range(4, 34):
        ws.cell(r, 1, r - 3)
        ws.cell(r, 6).value = f'=IF(C{r}="","",(C{r}*D{r})+E{r})'
        ws.cell(r, 7).value = f'=IF(C{r}="","",ROUND(SQRT(2*C{r}*365*2/0.5),0))'
        ws.cell(r, 8).value = f'=IF(C{r}="","",ROUND(G{r}/C{r},0))'
        style_data_row(ws, r, len(headers), LIGHT_GRAY if (r - 3) % 2 == 0 else None)

    # Parámetros
    ws2 = wb.create_sheet("Parámetros")
    ws2.sheet_properties.tabColor = "78909C"
    ws2.cell(1, 1, "PARÁMETROS DE REFERENCIA").font = Font(bold=True, size=14, color=GOLD)
    ws2.cell(3, 1, "Parámetro").font = Font(bold=True)
    ws2.cell(3, 2, "Valor Típico").font = Font(bold=True)
    ws2.cell(3, 3, "Notas").font = Font(bold=True)
    params = [
        ('Coste de pedido (€)', '2-5 €', 'Coste admin de hacer un pedido'),
        ('Coste almacenamiento (%)', '20-30% anual', '% del valor del producto'),
        ('Lead time cárnicos', '1-2 días', 'Entregas diarias habituales'),
        ('Lead time pescado', '1 día', 'Entrega diaria obligatoria'),
        ('Lead time secos', '3-5 días', 'Pedidos semanales'),
        ('Lead time congelados', '2-3 días', 'Pedidos 2x/semana'),
        ('Stock seguridad (días)', '1-3 días', 'Según criticidad del producto'),
    ]
    for i, (p, v, n) in enumerate(params, 1):
        r = i + 3
        ws2.cell(r, 1, p)
        ws2.cell(r, 2, v)
        ws2.cell(r, 3, n)
    for c in range(1, 4):
        ws2.column_dimensions[get_column_letter(c)].width = 30

    wb.save(os.path.join(OUT, 'BONUS-09-calculadora-punto-pedido.xlsx'))
    print("✅ BONUS-09-calculadora-punto-pedido.xlsx")


# ── Run all ──
if __name__ == '__main__':
    print(f"\n📦 Generando Kit Control de Inventario y Compras en {OUT}\n")
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_bonus_08()
    gen_bonus_09()
    print(f"\n✅ 9 archivos generados en {OUT}\n")
