#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Catering / Eventos.
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-catering"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

PRODUCCION_COLOR = "FFF3E0"
TRANSPORTE_COLOR = "E3F2FD"
MONTAJE_COLOR = "F3E5F5"
SERVICIO_COLOR = "E8F5E9"
DESMONTAJE_COLOR = "FCE4EC"
ADMIN_COLOR = "FFF8E1"
EVENTO_COLOR = "E0F2F1"
LIMPIEZA_COLOR = "EFEBE9"

ZONE_COLORS = {
    "Producción": PRODUCCION_COLOR,
    "Transporte": TRANSPORTE_COLOR,
    "Montaje": MONTAJE_COLOR,
    "Servicio": SERVICIO_COLOR,
    "Desmontaje": DESMONTAJE_COLOR,
    "Admin": ADMIN_COLOR,
    "Evento": EVENTO_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "General": LIGHT_GRAY,
}

title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Catering / Eventos"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for section_title, tasks in sections:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for idx, (task, zone) in enumerate(tasks, 1):
            zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")

            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=zone).font = data_font
            ws.cell(row=row, column=3).fill = zone_fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align

            check_cell = ws.cell(row=row, column=5)
            check_cell.font = checkbox_font
            check_cell.alignment = center_align
            check_cell.border = thin_border
            dv.add(check_cell)

            ws.cell(row=row, column=6).fill = input_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


# ═══════════════════════════════════════════════════════════
# 01 — PREPARACIÓN Y PRODUCCIÓN OFF-SITE
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Preparación y Producción Off-Site", [
        "Cómo usar esta plantilla:",
        "▸ Este checklist cubre la producción en cocina central antes del evento.",
        "▸ Incluye mise en place, elaboraciones, emplatado previo y empaquetado.",
        "▸ Marca ✓ cada tarea completada. Adapta tiempos a tu operación.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables: responsable, hora, notas.",
        "▸ Añade o elimina tareas según el tipo de evento.",
        "▸ Imprime en A4 y cuelga en la zona de producción.",
    ])
    create_task_sheet(wb, "Producción", PRODUCCION_COLOR, "Preparación y Producción Off-Site", [
        ("Planificación previa (48-72h antes)", [
            ("Revisar ficha técnica del evento: nº comensales, menú, alérgenos", "Admin"),
            ("Confirmar menú definitivo con cliente y restricciones dietéticas", "Admin"),
            ("Generar lista de compras por proveedor y categoría", "Admin"),
            ("Verificar stock de materia prima vs. necesidades del evento", "Producción"),
            ("Realizar pedidos a proveedores con margen de seguridad (+10%)", "Admin"),
            ("Confirmar logística de transporte: vehículo, hora, ruta", "Transporte"),
            ("Preparar timeline de producción con horas de inicio por partida", "Producción"),
            ("Asignar roles y responsables por partida/estación", "Admin"),
        ]),
        ("Producción D-1 (día anterior)", [
            ("Recepción y control de mercancía: temperaturas, cantidades, calidad", "Producción"),
            ("Almacenar correctamente: refrigeración, congelación, secos", "Producción"),
            ("Mise en place de salsas base, fondos y reducciones", "Producción"),
            ("Preparar marinados, adobos y maceraciones que requieran tiempo", "Producción"),
            ("Elaborar masas, bases de tartas y postres que necesiten reposo", "Producción"),
            ("Preparar pre-cocciones: confitados, braseados, sous-vide", "Producción"),
            ("Cortar y porcionar proteínas según ficha técnica", "Producción"),
            ("Preparar guarniciones y vegetales: lavar, cortar, blanquear", "Producción"),
            ("Etiquetar todo con producto, fecha, lote y alérgenos", "Producción"),
            ("Organizar cámaras por orden de carga para transporte", "Producción"),
        ]),
        ("Producción día D (día del evento)", [
            ("Verificar temperaturas de cámaras y producto almacenado", "Producción"),
            ("Completar elaboraciones finales: hornear, grillar, freír", "Producción"),
            ("Preparar salsas y aderezos de última hora", "Producción"),
            ("Emplatado previo de entrantes fríos y canapés", "Producción"),
            ("Montar bandejas de presentación y centros de mesa comestibles", "Producción"),
            ("Verificar cantidades vs. ficha del evento (check cruzado)", "Producción"),
            ("Empaquetado en GN y contenedores isotérmicos etiquetados", "Producción"),
            ("Control de temperatura pre-transporte (registrar en APPCC)", "Producción"),
            ("Verificar utensilios, menaje y material de servicio empaquetado", "Producción"),
            ("Limpieza y desinfección de la zona de producción post-carga", "Limpieza"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx"))


# ═══════════════════════════════════════════════════════════
# 02 — TRANSPORTE Y LOGÍSTICA
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Transporte y Logística", [
        "Cómo usar esta plantilla:",
        "▸ Checklist para el transporte seguro de alimentos y material.",
        "▸ Cubre carga, control de temperatura, descarga y trazabilidad.",
        "▸ Cumple normativa de transporte de alimentos refrigerados.",
        "",
        "Personalización:",
        "▸ Adapta a tu tipo de vehículo (furgoneta isotérmica, camión).",
        "▸ Añade rutas habituales y tiempos estimados.",
    ])
    create_task_sheet(wb, "Transporte", TRANSPORTE_COLOR, "Transporte y Logística", [
        ("Pre-carga del vehículo", [
            ("Verificar limpieza y desinfección del vehículo de transporte", "Transporte"),
            ("Comprobar funcionamiento del sistema de refrigeración del vehículo", "Transporte"),
            ("Pre-enfriar vehículo a temperatura objetivo (≤5°C frío, ≤-18°C congelado)", "Transporte"),
            ("Verificar combustible, documentación y seguro del vehículo", "Transporte"),
            ("Preparar kit de emergencia: termómetro, mantas térmicas, bolsas isotérmicas", "Transporte"),
            ("Cargar material de servicio primero (más pesado al fondo)", "Transporte"),
        ]),
        ("Carga de alimentos", [
            ("Registrar temperatura de cada GN/contenedor antes de cargar", "Transporte"),
            ("Cargar alimentos congelados primero, refrigerados después", "Transporte"),
            ("Separar alimentos crudos de elaborados (evitar contaminación cruzada)", "Transporte"),
            ("Asegurar contenedores para evitar movimiento durante el transporte", "Transporte"),
            ("Verificar listado de carga vs. packing list del evento", "Transporte"),
            ("Cargar utensilios de servicio, menaje y decoración", "Transporte"),
            ("Cerrar y sellar vehículo. Registrar hora de salida", "Transporte"),
        ]),
        ("Durante el transporte", [
            ("Monitorizar temperatura cada 30 min (registrar en hoja APPCC)", "Transporte"),
            ("Respetar cadena de frío: máx. 2h de transporte refrigerado", "Transporte"),
            ("Comunicar ETA al equipo de montaje en destino", "Transporte"),
            ("En caso de avería: activar protocolo de emergencia (bolsas isotérmicas)", "Transporte"),
        ]),
        ("Descarga en venue", [
            ("Registrar hora de llegada y temperatura del vehículo", "Transporte"),
            ("Descargar alimentos refrigerados primero al punto frío del venue", "Transporte"),
            ("Verificar estado visual de los alimentos post-transporte", "Transporte"),
            ("Registrar temperatura de producto al descargar (APPCC)", "Transporte"),
            ("Descargar material de servicio y menaje en zona de montaje", "Transporte"),
            ("Verificar listado de descarga vs. packing list (check cruzado)", "Transporte"),
            ("Aparcar vehículo y dejar accesible para desmontaje posterior", "Transporte"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "02-partidas-cocina.xlsx"))


# ═══════════════════════════════════════════════════════════
# 03 — TAREAS DEL DIRECTOR / EVENT MANAGER
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Tareas del Director / Event Manager", [
        "Cómo usar esta plantilla:",
        "▸ Checklist completo para el director de catering o event manager.",
        "▸ Cubre desde la reunión inicial con el cliente hasta el post-evento.",
        "▸ Incluye coordinación de equipos, timing y control de calidad.",
        "",
        "Personalización:",
        "▸ Adapta los timings a tu tipo de evento habitual.",
        "▸ Usa la sección de notas para incidencias.",
    ])
    create_task_sheet(wb, "Event Manager", ADMIN_COLOR, "Tareas del Director / Event Manager", [
        ("Semana del evento (D-7 a D-2)", [
            ("Reunión de producción con chef: confirmar menú, cantidades, timing", "Admin"),
            ("Confirmar personal: nº camareros, cocineros, barman, runners", "Admin"),
            ("Enviar briefing al equipo: hora llegada, dress code, contacto", "Admin"),
            ("Confirmar con venue: acceso carga, horarios, instalaciones", "Admin"),
            ("Verificar material: menaje, cristalería, mantelería, decoración", "Admin"),
            ("Preparar plano de sala/montaje con distribución de mesas y estaciones", "Admin"),
            ("Confirmar proveedores externos: florista, DJ, fotógrafo, AV", "Admin"),
            ("Preparar timing sheet detallado del evento (minuto a minuto)", "Admin"),
        ]),
        ("Día del evento — Pre-servicio", [
            ("Llegar al venue mínimo 3h antes del inicio del evento", "Admin"),
            ("Verificar acceso, electricidad, agua, baños y punto frío", "Admin"),
            ("Coordinar descarga y ubicación del equipo de cocina", "Admin"),
            ("Supervisar montaje de sala: mesas, sillas, mantelería", "Montaje"),
            ("Verificar montaje de buffet/estaciones según plano", "Montaje"),
            ("Briefing con equipo de sala: timing, menú, alérgenos, VIPs", "Admin"),
            ("Briefing con equipo de cocina: orden de salida, temperaturas", "Admin"),
            ("Contacto final con cliente: confirmar cambios de última hora", "Admin"),
            ("Verificar que música, iluminación y temperatura de sala son correctos", "Admin"),
        ]),
        ("Durante el evento", [
            ("Controlar timing de servicio: entrada, platos, postre, café", "Admin"),
            ("Supervisar calidad de emplatado y temperatura de los platos", "Admin"),
            ("Coordinar reposición de buffet y estaciones", "Servicio"),
            ("Gestionar incidencias: alérgenos no previstos, cambios de última hora", "Admin"),
            ("Mantener comunicación con cliente y wedding planner/organizador", "Admin"),
            ("Supervisar servicio de barra y bebidas", "Servicio"),
            ("Controlar tiempos de recogida de platos entre tiempos", "Servicio"),
            ("Verificar satisfacción de comensales (ronda discreta)", "Admin"),
        ]),
        ("Post-evento", [
            ("Supervisar desmontaje ordenado y limpieza del venue", "Desmontaje"),
            ("Verificar inventario de material recogido vs. entregado", "Admin"),
            ("Registrar incidencias, roturas y mermas", "Admin"),
            ("Confirmar con venue que el espacio queda en condiciones", "Admin"),
            ("Enviar agradecimiento al cliente (24-48h post-evento)", "Admin"),
            ("Reunión de retrospectiva con equipo (feedback y mejoras)", "Admin"),
            ("Actualizar P&L del evento: costes reales vs. presupuestado", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "03-tareas-manager.xlsx"))


# ═══════════════════════════════════════════════════════════
# 04 — TAREAS POR PERFIL / ROL
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Tareas por Perfil / Rol", [
        "Cómo usar esta plantilla:",
        "▸ Checklists separados por rol del equipo de catering.",
        "▸ Cada perfil sabe exactamente qué hacer antes, durante y después.",
        "▸ Ideal para onboarding de personal eventual/extra.",
        "",
        "Personalización:",
        "▸ Ajusta los perfiles a tu estructura de equipo.",
        "▸ Imprime solo la pestaña del perfil que necesites.",
    ])
    create_task_sheet(wb, "Chef de Evento", PRODUCCION_COLOR, "Chef de Evento / Jefe de Cocina", [
        ("Pre-servicio", [
            ("Verificar toda la producción recibida: cantidades, temperaturas, calidad", "Producción"),
            ("Organizar mise en place de la cocina del venue", "Producción"),
            ("Montar estaciones de trabajo: caliente, frío, postres", "Producción"),
            ("Verificar equipamiento de cocina del venue: hornos, fuegos, frío", "Producción"),
            ("Realizar prueba de emplatado con plato testigo", "Producción"),
            ("Briefing con equipo de cocina: orden de salida, timing, alérgenos", "Producción"),
        ]),
        ("Durante el servicio", [
            ("Controlar orden de salida de platos según timing", "Producción"),
            ("Supervisar emplatado y calidad de cada pase", "Producción"),
            ("Controlar temperaturas de producto en línea (>65°C caliente, <5°C frío)", "Producción"),
            ("Coordinar con maître/event manager los tiempos de servicio", "Producción"),
            ("Gestionar adaptaciones: platos para alérgicos, menús especiales", "Producción"),
            ("Controlar mermas y sobrantes durante el servicio", "Producción"),
        ]),
        ("Post-servicio", [
            ("Supervisar limpieza de cocina y equipamiento", "Limpieza"),
            ("Inventariar sobrantes: decidir qué se recupera y qué se descarta", "Producción"),
            ("Registrar mermas y notas para mejorar escandallo futuro", "Producción"),
            ("Verificar que la cocina del venue queda en condiciones", "Limpieza"),
        ]),
    ])
    create_task_sheet(wb, "Maître - Jefe Sala", SERVICIO_COLOR, "Maître / Jefe de Sala", [
        ("Pre-servicio", [
            ("Verificar montaje de sala según plano: mesas, sillas, mantelería", "Montaje"),
            ("Comprobar cristalería, cubertería y vajilla en cada puesto", "Montaje"),
            ("Verificar centros de mesa, decoración y señalización", "Montaje"),
            ("Asignar secciones/mesas a cada camarero", "Servicio"),
            ("Briefing de sala: menú, maridaje, alérgenos, VIPs, protocolo", "Servicio"),
            ("Verificar mise en place de estación de camareros (office)", "Servicio"),
        ]),
        ("Durante el servicio", [
            ("Coordinar con cocina los tiempos de pase", "Servicio"),
            ("Supervisar servicio de mesa: técnica, presentación, actitud", "Servicio"),
            ("Gestionar peticiones especiales de comensales", "Servicio"),
            ("Controlar servicio de bebidas y maridaje", "Servicio"),
            ("Coordinar recogida de platos entre tiempos", "Servicio"),
            ("Mantener comunicación con event manager sobre timing", "Servicio"),
        ]),
        ("Post-servicio", [
            ("Supervisar recogida de sala: cristalería, cubertería, vajilla", "Desmontaje"),
            ("Contar y verificar menaje recogido vs. inventario inicial", "Desmontaje"),
            ("Gestionar propinas y feedback del equipo de sala", "Admin"),
            ("Reportar roturas e incidencias al event manager", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Camarero", SERVICIO_COLOR, "Camarero / Runner de Evento", [
        ("Pre-servicio", [
            ("Llegar con uniforme completo y limpio al venue", "Servicio"),
            ("Atender briefing: menú, alérgenos, VIPs, timing", "Servicio"),
            ("Montar y verificar mesa/sección asignada", "Montaje"),
            ("Pulir cristalería y cubertería de tu sección", "Montaje"),
            ("Verificar mise en place de tu estación (pan, agua, condimentos)", "Servicio"),
            ("Conocer ubicación de baños, salida emergencia, zona fumadores", "Servicio"),
        ]),
        ("Durante el servicio", [
            ("Servir agua y pan en mesa antes de la llegada de los comensales", "Servicio"),
            ("Servir platos siguiendo indicaciones del maître (lado, técnica)", "Servicio"),
            ("Retirar platos usados de forma discreta y eficiente", "Servicio"),
            ("Reponer agua, pan y vino según necesidad", "Servicio"),
            ("Atender peticiones de comensales de forma proactiva", "Servicio"),
            ("Comunicar al maître cualquier incidencia o petición especial", "Servicio"),
        ]),
        ("Post-servicio", [
            ("Recoger tu sección: mantelería, cristalería, cubertería", "Desmontaje"),
            ("Separar menaje sucio por tipo en zona de lavado/acopio", "Desmontaje"),
            ("Ayudar en desmontaje general si se requiere", "Desmontaje"),
            ("Reportar roturas de tu sección al maître", "Desmontaje"),
        ]),
    ])
    create_task_sheet(wb, "Barman", SERVICIO_COLOR, "Barman / Responsable de Barra", [
        ("Pre-servicio", [
            ("Montar barra según plano: frontal, back-bar, zona de preparación", "Montaje"),
            ("Verificar stock de bebidas: vinos, cerveza, spirits, refrescos, agua", "Servicio"),
            ("Preparar hielo suficiente y verificar máquina de hielo o cubetas", "Servicio"),
            ("Mise en place de coctelería: cítricos, jarabes, garnish", "Servicio"),
            ("Verificar cristalería: copas de vino, vasos, flautas de champagne", "Servicio"),
            ("Preparar batches de cocktails si aplica (sangría, gin tonic premix)", "Servicio"),
        ]),
        ("Durante el servicio", [
            ("Gestionar servicio de bienvenida: cava, cocktail de bienvenida", "Servicio"),
            ("Servir vino en mesa si no hay sommelier (coordinar con camareros)", "Servicio"),
            ("Atender barra libre según horarios pactados", "Servicio"),
            ("Controlar consumo y reponer stock según ritmo del evento", "Servicio"),
            ("Mantener barra limpia y ordenada durante todo el evento", "Servicio"),
            ("Preparar cocktails especiales según carta del evento", "Servicio"),
        ]),
        ("Post-servicio", [
            ("Inventariar botellas sobrantes (abiertas y cerradas)", "Desmontaje"),
            ("Recoger y limpiar barra: cristalería, utensilios, superficie", "Desmontaje"),
            ("Empaquetar cristalería y material de barra para transporte", "Desmontaje"),
            ("Reportar consumo real vs. estimado al event manager", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "04-tareas-perfiles.xlsx"))


# ═══════════════════════════════════════════════════════════
# 05 — MONTAJE Y DESMONTAJE DEL VENUE
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Montaje y Desmontaje del Venue", [
        "Cómo usar esta plantilla:",
        "▸ Checklist detallado para montar y desmontar el espacio del evento.",
        "▸ Cubre sala, cocina temporal, barra, office y zonas auxiliares.",
        "▸ Incluye checklist de verificación pre-apertura de puertas.",
        "",
        "Personalización:",
        "▸ Adapta según tipo de venue: hotel, finca, espacio exterior, salón.",
        "▸ Añade zonas específicas de tu operación habitual.",
    ])
    create_task_sheet(wb, "Montaje", MONTAJE_COLOR, "Montaje del Venue", [
        ("Montaje de cocina temporal / office", [
            ("Ubicar cocina temporal según plano: acceso a electricidad y agua", "Montaje"),
            ("Montar mesas de trabajo, estantes y superficies de apoyo", "Montaje"),
            ("Instalar equipamiento: hornos portátiles, infernillos, calentadores", "Montaje"),
            ("Verificar punto de agua y desagüe para lavado", "Montaje"),
            ("Montar zona de emplatado con espacio suficiente para el equipo", "Montaje"),
            ("Instalar nevera/congelador portátil o verificar el del venue", "Montaje"),
            ("Organizar almacenamiento temporal: GN, contenedores, materia prima", "Montaje"),
            ("Colocar cubo de basura con bolsa y separación de residuos", "Montaje"),
        ]),
        ("Montaje de sala", [
            ("Distribuir mesas según plano de sala aprobado", "Montaje"),
            ("Colocar sillas y verificar estabilidad", "Montaje"),
            ("Poner mantelería: manteles, caminos, servilletas", "Montaje"),
            ("Montar cubertería, cristalería y vajilla en cada puesto", "Montaje"),
            ("Colocar centros de mesa y decoración floral", "Montaje"),
            ("Poner número/nombre de mesa y tarjetas de asiento si aplica", "Montaje"),
            ("Verificar iluminación de sala y ajustar a la ambientación", "Montaje"),
            ("Comprobar sonido/música y micrófonos si hay discursos", "Montaje"),
        ]),
        ("Montaje de buffet/estaciones", [
            ("Montar mesas de buffet según plano con niveles y decoración", "Montaje"),
            ("Colocar chafing dishes/calientaplatos y verificar combustible", "Montaje"),
            ("Ubicar estaciones de acción: cocina en vivo, wok, parrilla", "Montaje"),
            ("Montar estación de postres con refrigeración si necesario", "Montaje"),
            ("Colocar señalización de platos con nombre y alérgenos", "Montaje"),
            ("Montar barra de bebidas según plano", "Montaje"),
        ]),
        ("Check pre-apertura", [
            ("Verificación final de mesa testigo (foto para estándar)", "Admin"),
            ("Comprobar temperatura de sala (20-22°C ideal)", "Admin"),
            ("Verificar baños: jabón, papel, ambientador, limpieza", "Limpieza"),
            ("Comprobar acceso de emergencia despejado", "Admin"),
            ("Foto general de sala montada (archivo y redes sociales)", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Desmontaje", DESMONTAJE_COLOR, "Desmontaje Post-Evento", [
        ("Desmontaje de sala", [
            ("Recoger mantelería y separar: limpia, sucia, dañada", "Desmontaje"),
            ("Recoger cristalería por tipo en cajas con separadores", "Desmontaje"),
            ("Recoger cubertería y contar piezas vs. inventario", "Desmontaje"),
            ("Recoger vajilla y apilar por tipo", "Desmontaje"),
            ("Desmontar centros de mesa y decoración", "Desmontaje"),
            ("Recoger mesas, sillas y devolverlas a posición original", "Desmontaje"),
        ]),
        ("Desmontaje de cocina", [
            ("Apagar todos los equipos: hornos, infernillos, calentadores", "Desmontaje"),
            ("Limpiar superficies de trabajo y equipamiento", "Limpieza"),
            ("Recoger utensilios y GN. Contar y verificar", "Desmontaje"),
            ("Gestionar sobrantes: lo recuperable a contenedores, lo descartable a basura", "Desmontaje"),
            ("Desmontar cocina temporal y equipo portátil", "Desmontaje"),
            ("Limpiar zona de cocina y dejar como estaba", "Limpieza"),
        ]),
        ("Carga y salida", [
            ("Cargar material de cocina y menaje en vehículo", "Transporte"),
            ("Verificar que no queda nada en el venue (recorrido final)", "Desmontaje"),
            ("Verificar con responsable del venue que el espacio queda correcto", "Admin"),
            ("Registrar hora de salida y estado del material", "Admin"),
            ("Cerrar vehículo y verificar temperatura para transporte de vuelta", "Transporte"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx"))


# ═══════════════════════════════════════════════════════════
# 06 — TIPOS DE EVENTO
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Tipos de Evento — Checklists Específicos", [
        "Cómo usar esta plantilla:",
        "▸ Checklists adaptados a cada tipo de evento de catering.",
        "▸ Bodas, corporativos, cocktails, banquetes formales y eventos outdoor.",
        "▸ Usa la pestaña del tipo de evento que tengas próximo.",
        "",
        "Personalización:",
        "▸ Cada evento tiene particularidades. Usa estas listas como base.",
        "▸ Combina con los checklists generales (01-05) para cobertura completa.",
    ])
    create_task_sheet(wb, "Bodas", EVENTO_COLOR, "Boda — Checklist Específico", [
        ("Pre-boda (semana antes)", [
            ("Confirmar nº comensales definitivo con wedding planner", "Admin"),
            ("Confirmar distribución de mesas y plano de asientos", "Admin"),
            ("Verificar menú infantil y nº de niños", "Admin"),
            ("Confirmar horario: cocktail, banquete, barra libre, tarta", "Admin"),
            ("Coordinar con florista, DJ, fotógrafo los accesos y timing", "Admin"),
            ("Preparar tarta nupcial o coordinar con pastelería externa", "Producción"),
        ]),
        ("Día de la boda", [
            ("Cocktail de bienvenida: timing, canapés, bebidas", "Servicio"),
            ("Servicio de mesa: coordinar pases con protocolo nupcial", "Servicio"),
            ("Corte de tarta: coordinar con fotógrafo y DJ", "Servicio"),
            ("Barra libre: controlar horario de inicio y fin", "Servicio"),
            ("Servicio de recena/snack (si aplica)", "Servicio"),
            ("Verificar que novios y mesa principal están atendidos en todo momento", "Servicio"),
        ]),
    ])
    create_task_sheet(wb, "Corporativo", ADMIN_COLOR, "Evento Corporativo — Checklist", [
        ("Pre-evento", [
            ("Confirmar con empresa: agenda del evento, pausas café, comidas", "Admin"),
            ("Verificar requisitos AV: proyector, pantalla, micro, podio", "Admin"),
            ("Preparar coffee breaks: mañana y tarde con variación", "Producción"),
            ("Confirmar restricciones dietéticas de asistentes", "Admin"),
            ("Coordinar con hotel/venue los accesos de montaje", "Admin"),
        ]),
        ("Durante el evento", [
            ("Coffee break mañana: servir a la hora exacta pactada", "Servicio"),
            ("Almuerzo/comida: servicio según formato (buffet, sentado, boxes)", "Servicio"),
            ("Coffee break tarde: reponer y variar oferta vs. mañana", "Servicio"),
            ("Mantener estación de agua, café y snacks disponible todo el día", "Servicio"),
            ("Cocktail networking: montar rápido tras sesión final", "Servicio"),
        ]),
    ])
    create_task_sheet(wb, "Cocktail Standing", SERVICIO_COLOR, "Cocktail / Standing — Checklist", [
        ("Preparación", [
            ("Calcular ratio: 12-15 piezas por persona para cocktail de 2h", "Producción"),
            ("Preparar bandejas de canapés con variedad: mar, tierra, vegetal", "Producción"),
            ("Preparar estaciones de acción: ceviche, tartar, wok", "Producción"),
            ("Montar barra de cocktails con 3-4 opciones + refrescos + agua", "Montaje"),
            ("Calcular personal: 1 camarero / 20 invitados para pase de bandejas", "Admin"),
        ]),
        ("Durante el cocktail", [
            ("Salida de bandejas cada 10-15 min con rotación de oferta", "Servicio"),
            ("Reponer estaciones de acción según ritmo de consumo", "Servicio"),
            ("Controlar flujo de bebidas: no dejar a nadie sin copa", "Servicio"),
            ("Recoger cristalería y servilletas usadas de forma continua", "Servicio"),
            ("Mantener zonas limpias y presentables durante todo el evento", "Limpieza"),
        ]),
    ])
    create_task_sheet(wb, "Outdoor - Exterior", EVENTO_COLOR, "Evento Outdoor / Exterior — Checklist", [
        ("Logística outdoor", [
            ("Verificar previsión meteorológica 72h, 24h y 2h antes", "Admin"),
            ("Confirmar plan B en caso de lluvia (carpa, interior, fecha alternativa)", "Admin"),
            ("Verificar acceso de vehículos al terreno y zona de descarga", "Transporte"),
            ("Comprobar suministro eléctrico: generador o acometida del venue", "Admin"),
            ("Verificar punto de agua potable más cercano", "Admin"),
            ("Instalar carpa/toldo para cocina y zona de servicio si es necesario", "Montaje"),
        ]),
        ("Seguridad alimentaria outdoor", [
            ("Proteger alimentos del sol directo y calor ambiental", "Producción"),
            ("Duplicar controles de temperatura (alimentos se calientan más rápido)", "Producción"),
            ("Usar vitrinas con tapa/film para proteger de insectos", "Servicio"),
            ("Reducir tiempo de exposición: reponer más frecuentemente, menos cantidad", "Servicio"),
            ("Tener hielo extra y contenedores isotérmicos como backup", "Producción"),
            ("Verificar iluminación si el evento se extiende a la noche", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "06-eventos-festivos.xlsx"))


# ═══════════════════════════════════════════════════════════
# 07 — PLANTILLA PERSONALIZABLE
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Plantilla Personalizable — Catering", [
        "Cómo usar esta plantilla:",
        "▸ 3 plantillas en blanco para crear tus propias listas de tareas.",
        "▸ Formato por fase del evento, por zona del venue o por perfil.",
        "▸ Usa como base para eventos con requisitos especiales.",
        "",
        "Personalización:",
        "▸ Rellena las celdas con tus tareas propias.",
        "▸ Copia la pestaña que más se ajuste a tu operación.",
    ])

    for sheet_info in [
        ("Por Fase", EVENTO_COLOR, "Plantilla por Fase del Evento"),
        ("Por Zona", MONTAJE_COLOR, "Plantilla por Zona del Venue"),
        ("Por Perfil", SERVICIO_COLOR, "Plantilla por Perfil / Rol"),
    ]:
        sheet_name, color, title = sheet_info
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = color
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15

        ws.merge_cells("A1:G1")
        ws["A1"].value = title
        ws["A1"].font = title_font

        ws.merge_cells("A2:G2")
        ws["A2"].value = "AI Chef Pro · aichef.pro — Escribe tus propias tareas"
        ws["A2"].font = subtitle_font

        headers = ["#", "Tarea", "Zona/Fase", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
        ws.add_data_validation(dv)

        for r in range(5, 55):
            ws.cell(row=r, column=1, value=r - 4).font = data_font
            ws.cell(row=r, column=1).alignment = center_align
            ws.cell(row=r, column=1).border = thin_border
            for c in range(2, 8):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                if c == 4 or c == 6:
                    cell.fill = input_fill
                if c == 5:
                    dv.add(cell)
                    cell.alignment = center_align

    wb.save(os.path.join(OUTPUT_DIR, "07-plantilla-personalizable.xlsx"))


# ═══════════════════════════════════════════════════════════
# BONUS 01 — BRIEFING PRE-EVENTO
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 01 · Briefing Pre-Evento", [
        "Cómo usar esta plantilla:",
        "▸ Plantilla para la reunión de briefing con todo el equipo.",
        "▸ Se realiza 30-60 min antes de abrir puertas.",
        "▸ Incluye info del evento, menú, alérgenos, VIPs y roles.",
        "",
        "Personalización:",
        "▸ Rellena antes de cada evento y reparte al equipo.",
        "▸ Imprime una copia para cocina y otra para sala.",
    ])
    create_task_sheet(wb, "Briefing", ADMIN_COLOR, "Briefing Pre-Evento", [
        ("Información del evento", [
            ("Tipo de evento: _______________ (boda/corporativo/cocktail/otro)", "Admin"),
            ("Cliente: _______________ Contacto: _______________", "Admin"),
            ("Nº comensales confirmados: ___ Nº mesas: ___ Nº personal: ___", "Admin"),
            ("Hora apertura puertas: ___ Hora inicio servicio: ___", "Admin"),
            ("Hora fin estimada: ___ Hora desmontaje máximo: ___", "Admin"),
        ]),
        ("Menú y alérgenos", [
            ("Repasar menú completo con equipo: entrantes, principales, postres", "Producción"),
            ("Confirmar platos especiales: vegano, sin gluten, kosher, halal", "Producción"),
            ("Alérgenos críticos en mesa: identificar comensales y platos alternativos", "Producción"),
            ("Maridaje: vino blanco con ___, tinto con ___, cava para ___", "Servicio"),
            ("Orden de servicio: 1.___  2.___  3.___  4.___  5.___", "Servicio"),
        ]),
        ("Equipo y roles", [
            ("Chef de evento: _______________ (responsable cocina)", "Admin"),
            ("Maître: _______________ (responsable sala)", "Admin"),
            ("Barman: _______________ (responsable barra)", "Admin"),
            ("Camareros de sección: asignar mesas/zonas", "Admin"),
            ("Runners: asignar a cocina o sala según necesidad", "Admin"),
            ("Contacto de emergencia del venue: _______________", "Admin"),
        ]),
        ("VIPs y protocolo", [
            ("Identificar mesa del anfitrión/novios/CEO", "Servicio"),
            ("Comensales VIP con atención especial: _______________", "Servicio"),
            ("Protocolo de discursos/brindis: hora estimada ___", "Servicio"),
            ("Restricciones del venue: hora de fin de música, nivel de ruido", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))


# ═══════════════════════════════════════════════════════════
# BONUS 02 — CALENDARIO ANUAL DE EVENTOS
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 02 · Calendario Anual de Temporada Alta", [
        "Cómo usar esta plantilla:",
        "▸ Las 20 fechas clave del año para empresas de catering.",
        "▸ Incluye temporada alta de bodas, corporativos y festividades.",
        "▸ Cada fecha incluye antelación de preparación recomendada.",
        "",
        "Personalización:",
        "▸ Añade tus fechas locales (ferias, fiestas patronales).",
        "▸ Usa para planificar personal eventual y compras anticipadas.",
    ])

    ws = wb.create_sheet(title="Calendario")
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Calendario Anual de Temporada Alta — Catering"
    ws["A1"].font = title_font

    ws.merge_cells("A2:F2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — 20 fechas clave para planificar tu año"
    ws["A2"].font = subtitle_font

    headers = ["#", "Fecha / Período", "Tipo", "Preparación Especial", "Antelación", "Notas"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    events = [
        ("Enero — Reyes / Roscón", "Festividad", "Menús especiales, roscón gigante para empresas", "3 semanas"),
        ("Febrero — San Valentín", "Festividad", "Cenas románticas, menú degustación para parejas", "4 semanas"),
        ("Marzo — Comuniones (inicio)", "Celebración", "Menús infantiles + adultos, decoración temática", "2-3 meses"),
        ("Abril — Semana Santa", "Festividad", "Menús sin carne, bacalao, potajes, postres tradicionales", "3 semanas"),
        ("Mayo — Temporada de bodas (inicio)", "Boda", "Alta demanda de personal eventual, reservar proveedores", "6-12 meses"),
        ("Mayo — Comuniones (pico)", "Celebración", "Máxima demanda: 2-3 eventos simultáneos posibles", "3-4 meses"),
        ("Junio — Bodas + Graduaciones", "Mixto", "Eventos outdoor, planes B por lluvia, carpas", "6-12 meses"),
        ("Junio — Fin de curso corporativo", "Corporativo", "Team building + catering, formato cocktail outdoor", "2 meses"),
        ("Julio-Agosto — Bodas de verano", "Boda", "Gestión del calor, menús frescos, más personal", "6-12 meses"),
        ("Septiembre — Vuelta al cole corporativa", "Corporativo", "Kick-offs, convenciones, coffee breaks masivos", "1-2 meses"),
        ("Septiembre — Bodas de otoño (inicio)", "Boda", "Menús de temporada otoñal, setas, caza", "6-12 meses"),
        ("Octubre — Halloween", "Festividad", "Eventos temáticos, decoración, cocktails especiales", "4 semanas"),
        ("Octubre — Cenas de empresa (inicio)", "Corporativo", "Presupuestos para diciembre, reservas anticipadas", "2-3 meses"),
        ("Noviembre — Black Friday gastro", "Comercial", "Ofertas de packs de catering para empresas", "4 semanas"),
        ("Diciembre — Cenas de Navidad empresa", "Corporativo", "Máxima demanda corporativa del año. Personal extra", "2-3 meses"),
        ("Diciembre — Comidas de Navidad familia", "Celebración", "Menús navideños, cordero, marisco, turrones", "4-6 semanas"),
        ("31 Dic — Nochevieja / Cotillón", "Festividad", "Menú premium, cotillón, barra libre, uvas", "2-3 meses"),
        ("Todo el año — Bautizos", "Celebración", "Formato cocktail o sentado, menú infantil + adulto", "1-2 meses"),
        ("Todo el año — Funerales / Tanatorio", "Celebración", "Servicio discreto, coffee + canapés, entrega rápida", "24-48h"),
        ("Todo el año — Presentaciones producto", "Corporativo", "Branding del cliente, finger food premium, AV", "3-6 semanas"),
    ]

    for idx, (date, tipo, prep, antelacion) in enumerate(events, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx).font = data_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=date).font = bold_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        tipo_colors = {
            "Festividad": "E0F2F1",
            "Boda": "FCE4EC",
            "Corporativo": "E3F2FD",
            "Celebración": "FFF3E0",
            "Mixto": "F3E5F5",
            "Comercial": "FFF8E1",
        }
        color = tipo_colors.get(tipo, LIGHT_GRAY)
        ws.cell(row=row, column=3, value=tipo).font = data_font
        ws.cell(row=row, column=3).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=prep).font = data_font
        ws.cell(row=row, column=4).alignment = left_align
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=antelacion).font = bold_font
        ws.cell(row=row, column=5).alignment = center_align
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).fill = input_fill

    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx"))


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Generando Kit de Tareas: Catering / Eventos...")
    gen_01()
    print("  ✓ 01 — Preparación y Producción Off-Site")
    gen_02()
    print("  ✓ 02 — Transporte y Logística")
    gen_03()
    print("  ✓ 03 — Tareas del Director / Event Manager")
    gen_04()
    print("  ✓ 04 — Tareas por Perfil / Rol")
    gen_05()
    print("  ✓ 05 — Montaje y Desmontaje del Venue")
    gen_06()
    print("  ✓ 06 — Tipos de Evento")
    gen_07()
    print("  ✓ 07 — Plantilla Personalizable")
    gen_bonus_01()
    print("  ✓ BONUS 01 — Briefing Pre-Evento")
    gen_bonus_02()
    print("  ✓ BONUS 02 — Calendario Anual de Temporada Alta")
    print(f"\n✅ 9 archivos generados en {OUTPUT_DIR}")
