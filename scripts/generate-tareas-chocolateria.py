#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Chocolatería / Obrador de Chocolate (9 archivos).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-chocolateria"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Zone colors for chocolatería
OBRADOR_COLOR = "FFF3E0"      # Orange light - production
VITRINA_COLOR = "E3F2FD"      # Blue light - display/service
CAMARA_COLOR = "E0F7FA"       # Cyan light - cold storage
MOSTRADOR_COLOR = "F3E5F5"    # Purple light - counter
LIMPIEZA_COLOR = "EFEBE9"     # Brown light - cleaning
ADMIN_COLOR = "FFF8E1"        # Amber light - admin
ALMACEN_COLOR = "E8F5E9"      # Green light - storage
PACKAGING_COLOR = "F1F8E9"    # Light green - packaging

ZONE_COLORS = {
    "Obrador": OBRADOR_COLOR,
    "Templado": OBRADOR_COLOR,
    "Moldeado": OBRADOR_COLOR,
    "Producción": OBRADOR_COLOR,
    "Envasado": OBRADOR_COLOR,
    "Vitrina": VITRINA_COLOR,
    "Cámara": CAMARA_COLOR,
    "Conservación": CAMARA_COLOR,
    "Mostrador": MOSTRADOR_COLOR,
    "Caja": MOSTRADOR_COLOR,
    "Servicio": MOSTRADOR_COLOR,
    "Tienda": MOSTRADOR_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Admin": ADMIN_COLOR,
    "Gestión": ADMIN_COLOR,
    "Almacén": ALMACEN_COLOR,
    "Packaging": PACKAGING_COLOR,
    "General": LIGHT_GRAY,
    "Equipo": ADMIN_COLOR,
    "RRSS": ADMIN_COLOR,
    "Logística": ALMACEN_COLOR,
}

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Chocolatería / Obrador de Chocolate"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for section_title, tasks in sections:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for idx, (task, zone) in enumerate(tasks, 1):
            zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")

            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=zone).font = data_font
            ws.cell(row=row, column=3).fill = zone_fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align

            check_cell = ws.cell(row=row, column=5)
            check_cell.font = checkbox_font
            check_cell.alignment = center_align
            check_cell.border = thin_border
            dv.add(check_cell)

            ws.cell(row=row, column=6).fill = input_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


def create_blank_template_sheet(wb, sheet_name, tab_color, title, col_headers):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Chocolatería / Obrador de Chocolate"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r in range(5, 30):
        for c in range(1, len(col_headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c == 1:
                cell.alignment = center_align
            elif c == 5:
                cell.font = checkbox_font
                cell.alignment = center_align
                dv.add(cell)
            elif c in (4, 6):
                cell.fill = input_fill
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.merge_cells(f"A31:G31")
    ws[f"A31"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A31"].font = small_font


def create_calendar_sheet(wb, sheet_name, tab_color, title, months_data):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:F2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Chocolatería / Obrador de Chocolate"
    ws["A2"].font = subtitle_font

    row = 4
    cal_headers = ["Mes", "Acciones clave / Fechas señaladas", "Productos destacados", "Temporada", "✓", "Notas"]
    for col_idx, h in enumerate(cal_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    ws.add_data_validation(dv)

    for month, actions, products, season in months_data:
        ws.cell(row=row, column=1, value=month).font = bold_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=actions).font = data_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=products).font = data_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = thin_border

        season_color = OBRADOR_COLOR if season == "Alta" else CAMARA_COLOR if season == "Baja" else VITRINA_COLOR
        ws.cell(row=row, column=4, value=season).font = data_font
        ws.cell(row=row, column=4).fill = PatternFill(start_color=season_color, end_color=season_color, fill_type="solid")
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border

        check = ws.cell(row=row, column=5)
        check.font = checkbox_font
        check.alignment = center_align
        check.border = thin_border
        dv.add(check)

        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = left_align

        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


# ═══════════════════════════════════════════════════════════
# 01 — APERTURA Y CIERRE DE CHOCOLATERÍA
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Apertura y Cierre de Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ Checklist completo para la apertura y cierre diario de tu chocolatería u obrador.",
        "▸ Incluye encendido de equipos de templado, vitrinas, montaje de servicio y cierre.",
        "▸ Marca ✓ cada tarea completada. Adapta horarios a tu negocio.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables: responsable, hora, notas.",
        "▸ Ajusta temperaturas según tu equipamiento y normativa local.",
        "▸ Imprime en A4 y deja en obrador y mostrador.",
    ])
    create_task_sheet(wb, "Apertura", OBRADOR_COLOR, "Apertura de Chocolatería", [
        ("Encendido y verificación de equipos", [
            ("Encender temperadora de chocolate y precalentar a 45°C", "Templado"),
            ("Verificar cámara de conservación a 15-18°C (bombones)", "Conservación"),
            ("Verificar nevera de rellenos y ganaches a 4°C", "Cámara"),
            ("Encender iluminación de vitrina y sala de tienda", "Vitrina"),
            ("Encender aire acondicionado — mantener sala 20-22°C", "Tienda"),
            ("Registrar temperaturas en hoja de control APPCC", "Admin"),
        ]),
        ("Montaje de servicio", [
            ("Montar vitrina: bombones, tabletas, figuras, trufas", "Vitrina"),
            ("Verificar etiquetado de alérgenos en cada producto expuesto", "Vitrina"),
            ("Reponer bolsas, cajas regalo, papel de seda, cintas", "Packaging"),
            ("Verificar stock de cajas de bombones surtidos", "Packaging"),
            ("Preparar zona de degustación con muestras del día", "Mostrador"),
            ("Activar TPV y verificar fondo de caja", "Caja"),
            ("Colocar pizarra con novedades y productos destacados", "Tienda"),
        ]),
        ("Verificaciones finales", [
            ("Comprobar que música ambiental funciona", "Tienda"),
            ("Verificar limpieza de cristales de vitrina", "Vitrina"),
            ("Confirmar pedidos online pendientes de preparar", "Admin"),
            ("Revisar agenda: pedidos especiales, eventos, encargos", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Cierre", LIMPIEZA_COLOR, "Cierre de Chocolatería", [
        ("Cierre de tienda", [
            ("Retirar productos de vitrina y guardar en conservación", "Vitrina"),
            ("Cubrir bombones que queden expuestos con film", "Vitrina"),
            ("Cuadre de caja y registro de ventas del día", "Caja"),
            ("Apagar TPV y cerrar terminal de pago", "Caja"),
            ("Limpiar mostradores y cristales de vitrina", "Mostrador"),
        ]),
        ("Cierre de obrador", [
            ("Limpiar y secar temperadora de chocolate", "Templado"),
            ("Limpiar moldes usados y dejar secando", "Moldeado"),
            ("Guardar coberturas abiertas en recipientes herméticos", "Almacén"),
            ("Verificar temperaturas de cámaras y registrar", "Conservación"),
            ("Limpiar superficies de mármol/granito de obrador", "Obrador"),
            ("Barrer y fregar suelo de obrador", "Limpieza"),
        ]),
        ("Seguridad y cierre", [
            ("Apagar equipos excepto cámaras y neveras", "General"),
            ("Verificar que gas, agua y luces quedan apagadas", "General"),
            ("Sacar basura y reciclar cartón/plástico", "Limpieza"),
            ("Activar alarma y cerrar con llave", "General"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "01-apertura-cierre.xlsx"))
    print("  ✓ 01-apertura-cierre.xlsx")


# ═══════════════════════════════════════════════════════════
# 02 — PARTIDAS DE PRODUCCIÓN
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Partidas de Producción de Chocolate", [
        "Cómo usar esta plantilla:",
        "▸ Control de producción artesanal: templado, moldeado, bombones, tabletas.",
        "▸ Registra lotes, temperaturas y tiempos de cada proceso.",
        "▸ Adapta las recetas base a tus fórmulas propias.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables: responsable, hora, notas.",
        "▸ Ajusta temperaturas de templado según tipo de cobertura.",
        "▸ Imprime y deja en obrador junto a la temperadora.",
    ])
    create_task_sheet(wb, "Templado", OBRADOR_COLOR, "Control de Templado", [
        ("Preparación de coberturas", [
            ("Verificar stock de cobertura negra (55-70% cacao)", "Almacén"),
            ("Verificar stock de cobertura con leche (35-40% cacao)", "Almacén"),
            ("Verificar stock de cobertura blanca (28-33% cacao)", "Almacén"),
            ("Pesar cobertura según orden de producción del día", "Obrador"),
            ("Trocear cobertura y colocar en temperadora", "Templado"),
        ]),
        ("Proceso de templado — Chocolate negro", [
            ("Fundir a 50-55°C — registrar temperatura", "Templado"),
            ("Enfriar a 27-28°C sobre mármol o por siembra", "Templado"),
            ("Recalentar a 31-32°C — punto de trabajo", "Templado"),
            ("Test de templado: brillo, snap, contracción en molde", "Templado"),
            ("Registrar lote, hora y temperatura final", "Admin"),
        ]),
        ("Proceso de templado — Chocolate con leche", [
            ("Fundir a 45-50°C — registrar temperatura", "Templado"),
            ("Enfriar a 26-27°C sobre mármol o por siembra", "Templado"),
            ("Recalentar a 29-30°C — punto de trabajo", "Templado"),
            ("Test de templado y registrar lote", "Admin"),
        ]),
        ("Proceso de templado — Chocolate blanco", [
            ("Fundir a 40-45°C — registrar temperatura", "Templado"),
            ("Enfriar a 25-26°C sobre mármol o por siembra", "Templado"),
            ("Recalentar a 27-28°C — punto de trabajo", "Templado"),
            ("Test de templado y registrar lote", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Moldeado", OBRADOR_COLOR, "Producción: Moldeado y Bombones", [
        ("Moldeado de bombones", [
            ("Preparar moldes limpios y a temperatura ambiente (20-22°C)", "Moldeado"),
            ("Verter chocolate templado en moldes — golpear para eliminar burbujas", "Moldeado"),
            ("Vaciar exceso — dejar cáscara uniforme de 2-3mm", "Moldeado"),
            ("Invertir moldes sobre rejilla y dejar escurrir", "Moldeado"),
            ("Enfriar en cámara a 12-14°C durante 10-15 min", "Conservación"),
            ("Preparar rellenos/ganaches según receta del día", "Obrador"),
            ("Rellenar bombones al 80% dejando margen para tapa", "Moldeado"),
            ("Sellar con chocolate templado — alisar con espátula", "Moldeado"),
            ("Enfriar en cámara 15-20 min hasta contracción", "Conservación"),
            ("Desmoldar, verificar brillo y almacenar en bandejas", "Moldeado"),
        ]),
        ("Producción de tabletas", [
            ("Preparar moldes de tableta limpios y atemperados", "Moldeado"),
            ("Verter chocolate templado y golpear/vibrar", "Moldeado"),
            ("Añadir inclusiones si aplica: frutos secos, frutas, sal", "Moldeado"),
            ("Alisar superficie con espátula", "Moldeado"),
            ("Enfriar en cámara a 12-14°C durante 20-30 min", "Conservación"),
            ("Desmoldar, verificar acabado y pesar cada tableta", "Moldeado"),
            ("Envolver en papel/aluminio y etiquetar con lote y fecha", "Envasado"),
        ]),
        ("Control de calidad y registro", [
            ("Verificar brillo, textura y snap de productos", "Obrador"),
            ("Registrar lotes de producción: producto, cantidad, fecha", "Admin"),
            ("Registrar mermas y chocolate descartado", "Admin"),
            ("Almacenar producto terminado a 15-18°C", "Conservación"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "02-partidas-produccion.xlsx"))
    print("  ✓ 02-partidas-produccion.xlsx")


# ═══════════════════════════════════════════════════════════
# 03 — TAREAS DEL MANAGER
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Tareas del Manager de Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ Tareas diarias, semanales y mensuales del encargado/director.",
        "▸ Incluye control de food cost, mermas y KPIs de producción.",
        "▸ Personaliza según el tamaño de tu operación.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables: responsable, hora, notas.",
        "▸ Adapta los KPIs a tus objetivos mensuales.",
    ])
    create_task_sheet(wb, "Diario", ADMIN_COLOR, "Tareas Diarias del Manager", [
        ("Primera hora (apertura)", [
            ("Revisar agenda: pedidos especiales, encargos, entregas", "Gestión"),
            ("Comprobar producción planificada vs stock actual", "Gestión"),
            ("Briefing con equipo: prioridades del día, encargos especiales", "Equipo"),
            ("Verificar temperaturas de obrador, cámaras y tienda", "Admin"),
            ("Revisar pedidos online recibidos overnight", "Admin"),
        ]),
        ("Durante el día", [
            ("Supervisar calidad de producción: templado, acabados, peso", "Producción"),
            ("Control de mermas y chocolate descartado", "Gestión"),
            ("Gestionar atención al cliente y consultas especiales", "Tienda"),
            ("Seguimiento de pedidos a proveedores pendientes", "Admin"),
            ("Publicar contenido en RRSS: producto del día, proceso", "RRSS"),
        ]),
        ("Cierre del día", [
            ("Revisar ventas del día vs objetivo", "Gestión"),
            ("Verificar stock para producción de mañana", "Almacén"),
            ("Anotar incidencias y notas para el día siguiente", "Admin"),
            ("Confirmar equipo y horarios de mañana", "Equipo"),
        ]),
    ])
    create_task_sheet(wb, "Semanal", ADMIN_COLOR, "Tareas Semanales del Manager", [
        ("Lunes — Planificación", [
            ("Planificar producción de la semana según demanda", "Gestión"),
            ("Revisar stock de coberturas, rellenos y packaging", "Almacén"),
            ("Lanzar pedidos a proveedores para la semana", "Logística"),
            ("Revisar pedidos especiales y encargos de la semana", "Admin"),
        ]),
        ("Martes — Calidad", [
            ("Auditar temperaturas registradas de la semana anterior", "Admin"),
            ("Verificar fechas de caducidad de rellenos y ganaches", "Conservación"),
            ("Revisar mermas acumuladas y causas", "Gestión"),
        ]),
        ("Miércoles — Equipo", [
            ("Revisar rendimiento individual: productividad, calidad", "Equipo"),
            ("Formación puntual: nueva técnica o producto", "Equipo"),
            ("Ajustar turnos si hay picos de demanda previstos", "Equipo"),
        ]),
        ("Jueves — Comercial", [
            ("Revisar ventas por producto: tabletas, bombones, figuras", "Gestión"),
            ("Actualizar precios si hay cambios en materia prima", "Admin"),
            ("Planificar ofertas o promociones del fin de semana", "RRSS"),
        ]),
        ("Viernes — Food Cost", [
            ("Calcular food cost semanal: materia prima / ventas", "Gestión"),
            ("Comparar food cost real vs objetivo (30-35%)", "Gestión"),
            ("Ajustar recetas o porciones si hay desviación", "Producción"),
            ("Preparar resumen semanal para dirección", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "Mensual", ADMIN_COLOR, "Tareas Mensuales del Manager", [
        ("Inventario y finanzas", [
            ("Inventario completo: coberturas, frutos secos, licores, packaging", "Almacén"),
            ("Calcular food cost mensual y comparar con objetivo", "Gestión"),
            ("Revisar facturas de proveedores y negociar si aplica", "Admin"),
            ("Análisis de ventas por categoría: bombones, tabletas, figuras, regalos", "Gestión"),
        ]),
        ("Equipo y formación", [
            ("Evaluación de equipo: feedback individual", "Equipo"),
            ("Planificar formación del mes siguiente", "Equipo"),
            ("Revisar y actualizar procedimientos operativos", "Admin"),
        ]),
        ("Producto y marketing", [
            ("Lanzar nuevo producto o sabor del mes", "Producción"),
            ("Actualizar carta y etiquetado si hay cambios", "Tienda"),
            ("Planificar contenido RRSS del mes siguiente", "RRSS"),
            ("Revisar Google Reviews y responder pendientes", "RRSS"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "03-tareas-manager.xlsx"))
    print("  ✓ 03-tareas-manager.xlsx")


# ═══════════════════════════════════════════════════════════
# 04 — TAREAS POR PERFIL
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Tareas por Perfil — Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ Checklists específicos por puesto de trabajo.",
        "▸ Cada perfil tiene sus tareas claramente definidas.",
        "▸ Ideal para onboarding de nuevos empleados.",
        "",
        "Perfiles incluidos:",
        "▸ Maestro chocolatero / Obrador",
        "▸ Dependiente de tienda",
        "▸ Encargado de turno",
    ])
    create_task_sheet(wb, "Chocolatero", OBRADOR_COLOR, "Maestro Chocolatero / Obrador", [
        ("Producción diaria", [
            ("Revisar orden de producción del día", "Obrador"),
            ("Preparar y pesar ingredientes según fichas técnicas", "Obrador"),
            ("Templar coberturas según tipo (negro/leche/blanco)", "Templado"),
            ("Producir bombones según recetas del día", "Moldeado"),
            ("Producir tabletas y figuras según planificación", "Moldeado"),
            ("Preparar ganaches y rellenos según ficha", "Obrador"),
            ("Envasar y etiquetar producto terminado", "Envasado"),
            ("Registrar lotes de producción y cantidades", "Admin"),
        ]),
        ("Calidad y control", [
            ("Verificar templado: brillo, snap, contracción", "Templado"),
            ("Controlar pesos de bombones y tabletas", "Obrador"),
            ("Registrar temperaturas de proceso", "Admin"),
            ("Registrar mermas y desperdicio", "Admin"),
        ]),
        ("Limpieza de obrador", [
            ("Limpiar temperadora al finalizar cada uso", "Templado"),
            ("Limpiar moldes y secar correctamente", "Moldeado"),
            ("Limpiar mármol de trabajo", "Obrador"),
            ("Dejar obrador ordenado para el turno siguiente", "Obrador"),
        ]),
    ])
    create_task_sheet(wb, "Dependiente", MOSTRADOR_COLOR, "Dependiente de Tienda", [
        ("Apertura de tienda", [
            ("Montar vitrina con productos del día", "Vitrina"),
            ("Verificar etiquetado de precios y alérgenos", "Vitrina"),
            ("Preparar zona de degustación", "Mostrador"),
            ("Activar TPV y comprobar fondo de caja", "Caja"),
            ("Reponer packaging: bolsas, cajas, papel de seda", "Packaging"),
        ]),
        ("Atención al cliente", [
            ("Atender clientes: asesorar sobre productos y maridajes", "Tienda"),
            ("Preparar cajas de bombones personalizadas", "Packaging"),
            ("Gestionar encargos especiales y pedidos por adelantado", "Admin"),
            ("Ofrecer degustación del producto destacado del día", "Mostrador"),
            ("Cobrar y emitir ticket/factura", "Caja"),
        ]),
        ("Mantenimiento de tienda", [
            ("Reponer vitrina conforme se venden productos", "Vitrina"),
            ("Mantener limpieza de mostradores y cristales", "Limpieza"),
            ("Vigilar temperatura de tienda (20-22°C ideal)", "Tienda"),
        ]),
        ("Cierre de tienda", [
            ("Cuadrar caja y registrar ventas", "Caja"),
            ("Retirar productos de vitrina para conservación", "Vitrina"),
            ("Limpiar mostradores, cristales y suelo de tienda", "Limpieza"),
        ]),
    ])
    create_task_sheet(wb, "Encargado", ADMIN_COLOR, "Encargado de Turno", [
        ("Supervisión general", [
            ("Verificar apertura correcta (obrador + tienda)", "General"),
            ("Asignar tareas al equipo según prioridades", "Equipo"),
            ("Supervisar calidad de producción", "Producción"),
            ("Gestionar pedidos urgentes de proveedores", "Logística"),
            ("Resolver incidencias con clientes", "Tienda"),
        ]),
        ("Control operativo", [
            ("Verificar stock mínimo de coberturas y packaging", "Almacén"),
            ("Supervisar temperaturas de cámaras y obrador", "Admin"),
            ("Control de mermas y desperdicio del turno", "Gestión"),
            ("Verificar cumplimiento de APPCC", "Admin"),
        ]),
        ("Cierre de turno", [
            ("Revisar producción completada vs planificada", "Gestión"),
            ("Dejar notas para el turno siguiente", "Admin"),
            ("Supervisar cierre correcto de obrador y tienda", "General"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "04-tareas-perfiles.xlsx"))
    print("  ✓ 04-tareas-perfiles.xlsx")


# ═══════════════════════════════════════════════════════════
# 05 — TAREAS SEMANALES Y MENSUALES
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Tareas Semanales y Mensuales — Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ Deep clean, mantenimiento de equipos, inventarios periódicos.",
        "▸ Programa estas tareas en el calendario semanal/mensual.",
        "▸ Marca ✓ cuando se complete cada tarea.",
        "",
        "Personalización:",
        "▸ Ajusta frecuencias según volumen de producción.",
        "▸ Añade tareas específicas de tu obrador.",
    ])
    create_task_sheet(wb, "Semanal", LIMPIEZA_COLOR, "Tareas Semanales", [
        ("Deep clean — Obrador", [
            ("Limpieza profunda de temperadora: desmontar y limpiar", "Templado"),
            ("Limpiar y desinfectar mármol de trabajo con producto alimentario", "Obrador"),
            ("Limpiar todos los moldes: verificar que no tengan restos", "Moldeado"),
            ("Desinfectar superficies de envasado", "Envasado"),
            ("Limpiar y organizar estanterías de obrador", "Obrador"),
            ("Fregar suelo a fondo con desengrasante alimentario", "Limpieza"),
        ]),
        ("Deep clean — Tienda", [
            ("Limpiar cristales de vitrina interior y exterior", "Vitrina"),
            ("Limpiar estanterías y expositores de tienda", "Tienda"),
            ("Limpiar escaparate", "Tienda"),
            ("Desinfectar zona de degustación", "Mostrador"),
        ]),
        ("Inventario semanal", [
            ("Contar stock de coberturas: negra, leche, blanca", "Almacén"),
            ("Contar stock de frutos secos y frutas deshidratadas", "Almacén"),
            ("Contar stock de rellenos y ganaches preparados", "Cámara"),
            ("Contar stock de packaging: cajas, bolsas, cintas, papel", "Packaging"),
            ("Contar stock de consumibles: guantes, film, etiquetas", "Almacén"),
            ("Anotar roturas de stock y generar pedido", "Admin"),
        ]),
        ("Mantenimiento de equipos", [
            ("Verificar funcionamiento de temperadora", "Templado"),
            ("Verificar temperaturas de cámaras frigoríficas", "Conservación"),
            ("Comprobar estado de moldes (rayados, deformados)", "Moldeado"),
            ("Verificar báscula de precisión con peso patrón", "Obrador"),
        ]),
    ])
    create_task_sheet(wb, "Mensual", ADMIN_COLOR, "Tareas Mensuales", [
        ("Mantenimiento profundo", [
            ("Mantenimiento profesional de temperadora si aplica", "Templado"),
            ("Limpieza de filtros de aire acondicionado", "General"),
            ("Revisión de instalación eléctrica y enchufes", "General"),
            ("Calibración de termómetros y sondas", "Obrador"),
            ("Verificar estado de extintores", "General"),
        ]),
        ("Inventario mensual completo", [
            ("Inventario completo: coberturas por tipo y proveedor", "Almacén"),
            ("Inventario: frutos secos, pralinés, licores, extractos", "Almacén"),
            ("Inventario: moldes, utillaje, herramientas", "Obrador"),
            ("Inventario: packaging completo (cajas, bolsas, lazos)", "Packaging"),
            ("Comparar inventario real vs teórico: detectar desfases", "Gestión"),
        ]),
        ("APPCC y seguridad", [
            ("Revisar registros APPCC del mes: temperaturas, limpieza", "Admin"),
            ("Verificar plan de control de plagas", "General"),
            ("Actualizar fichas técnicas si hay nuevos productos", "Admin"),
            ("Revisar etiquetado: ingredientes, alérgenos, lotes", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "05-tareas-semanales-mensuales.xlsx"))
    print("  ✓ 05-tareas-semanales-mensuales.xlsx")


# ═══════════════════════════════════════════════════════════
# 06 — EVENTOS Y TEMPORADA
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Eventos y Temporada — Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ Planificación de campañas estacionales y eventos especiales.",
        "▸ Navidad, San Valentín, Pascua, Día de la Madre y más.",
        "▸ Cada campaña tiene producción, packaging y marketing.",
        "",
        "Personalización:",
        "▸ Ajusta fechas y productos a tu catálogo.",
        "▸ Añade ferias locales y eventos de tu zona.",
    ])
    create_task_sheet(wb, "Navidad", OBRADOR_COLOR, "Campaña de Navidad (Nov-Dic)", [
        ("Preparación (Octubre)", [
            ("Diseñar catálogo navideño: turrones, bombones, figuras", "Gestión"),
            ("Pedir coberturas y materias primas extra", "Logística"),
            ("Diseñar y pedir packaging navideño: cajas, lazos, tarjetas", "Packaging"),
            ("Planificar producción extra: turnos, personal temporal", "Equipo"),
            ("Lanzar preventa en web y RRSS", "RRSS"),
        ]),
        ("Producción navideña (Nov-Dic)", [
            ("Producir turrones artesanales: praliné, guirlache, trufa", "Producción"),
            ("Producir figuras de chocolate: Papá Noel, árboles, bolas", "Moldeado"),
            ("Producir bombones especiales navideños", "Moldeado"),
            ("Producir cestas y lotes corporativos", "Envasado"),
            ("Control de calidad extra: cada lote verificado", "Admin"),
        ]),
        ("Tienda y marketing", [
            ("Decorar tienda y escaparate con temática navideña", "Tienda"),
            ("Montar expositor especial de regalos corporativos", "Vitrina"),
            ("Publicar catálogo navideño en RRSS y web", "RRSS"),
            ("Gestionar pedidos corporativos: presupuestos y entregas", "Admin"),
        ]),
    ])
    create_task_sheet(wb, "San Valentín", MOSTRADOR_COLOR, "Campaña de San Valentín (Feb)", [
        ("Preparación (Enero)", [
            ("Diseñar colección San Valentín: corazones, bombones premium", "Gestión"),
            ("Pedir moldes especiales si se necesitan", "Logística"),
            ("Diseñar packaging romántico: cajas corazón, cintas rojas", "Packaging"),
            ("Planificar promoción en RRSS: countdown, sorteos", "RRSS"),
        ]),
        ("Producción (1-14 Feb)", [
            ("Producir bombones en forma de corazón", "Moldeado"),
            ("Producir tabletas con mensaje grabado", "Moldeado"),
            ("Producir trufas y bombones premium edición limitada", "Producción"),
            ("Preparar cajas surtidas de bombones para regalo", "Envasado"),
        ]),
        ("Ventas y marketing", [
            ("Decorar tienda y vitrina con temática San Valentín", "Tienda"),
            ("Ofrecer servicio de personalización: mensaje en tableta", "Mostrador"),
            ("Publicar contenido diario en RRSS: countdown", "RRSS"),
            ("Gestionar pedidos especiales y entregas a domicilio", "Logística"),
        ]),
    ])
    create_task_sheet(wb, "Pascua", OBRADOR_COLOR, "Campaña de Pascua (Mar-Abr)", [
        ("Preparación (Febrero)", [
            ("Diseñar colección Pascua: huevos, figuras, monas", "Gestión"),
            ("Pedir moldes de huevos y figuras", "Logística"),
            ("Diseñar packaging primaveral", "Packaging"),
            ("Planificar producción de monas de chocolate si aplica", "Producción"),
        ]),
        ("Producción de Pascua", [
            ("Producir huevos de chocolate: pequeños, medianos, grandes", "Moldeado"),
            ("Producir figuras de Pascua: conejos, pollitos", "Moldeado"),
            ("Producir monas de chocolate decoradas", "Producción"),
            ("Producir bombones primaverales", "Moldeado"),
            ("Envasar y etiquetar colección de Pascua", "Envasado"),
        ]),
        ("Tienda y marketing", [
            ("Decorar tienda con temática primaveral/Pascua", "Tienda"),
            ("Montar escaparate con huevos y figuras", "Vitrina"),
            ("Publicar colección en RRSS con fotos de calidad", "RRSS"),
            ("Organizar taller de decoración de huevos si aplica", "Tienda"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "06-eventos-temporada.xlsx"))
    print("  ✓ 06-eventos-temporada.xlsx")


# ═══════════════════════════════════════════════════════════
# 07 — PLANTILLA PERSONALIZABLE
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Plantilla Personalizable — Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ 3 plantillas en blanco para crear tus propias listas de tareas.",
        "▸ Por zona, por turno y por perfil.",
        "▸ Ideal para tareas específicas de tu negocio.",
        "",
        "Personalización:",
        "▸ Escribe tus tareas en las filas vacías.",
        "▸ Usa los mismos códigos de zona que en las otras plantillas.",
    ])
    col_headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
    create_blank_template_sheet(wb, "Por Zona", OBRADOR_COLOR, "Plantilla Personalizable — Por Zona", col_headers)
    create_blank_template_sheet(wb, "Por Turno", MOSTRADOR_COLOR, "Plantilla Personalizable — Por Turno", col_headers)
    create_blank_template_sheet(wb, "Por Perfil", ADMIN_COLOR, "Plantilla Personalizable — Por Perfil", col_headers)
    wb.save(os.path.join(OUTPUT_DIR, "07-plantilla-personalizable.xlsx"))
    print("  ✓ 07-plantilla-personalizable.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS 01 — BRIEFING DIARIO
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 01 · Briefing Diario de Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ Reunión de 5 minutos al inicio del turno.",
        "▸ Producción del día, pedidos especiales, equipo, promociones.",
        "▸ Imprime y rellena cada mañana con el equipo.",
        "",
        "Personalización:",
        "▸ Las celdas verdes son editables.",
        "▸ Adapta las secciones a tu operación.",
    ])
    create_task_sheet(wb, "Briefing", ADMIN_COLOR, "Briefing Diario de Chocolatería", [
        ("Producción del día", [
            ("Productos a producir hoy: ___________________________", "Producción"),
            ("Cantidad de bombones planificada: _______ unidades", "Producción"),
            ("Cantidad de tabletas planificada: _______ unidades", "Producción"),
            ("Figuras/piezas especiales: ___________________________", "Producción"),
            ("Coberturas necesarias: negra ___ kg / leche ___ kg / blanca ___ kg", "Almacén"),
        ]),
        ("Pedidos especiales", [
            ("Pedido 1: _____________ — Entrega: ___/___", "Admin"),
            ("Pedido 2: _____________ — Entrega: ___/___", "Admin"),
            ("Pedido 3: _____________ — Entrega: ___/___", "Admin"),
            ("Pedidos online pendientes: ___________________________", "Admin"),
        ]),
        ("Equipo del turno", [
            ("Obrador: ___________________________", "Equipo"),
            ("Tienda: ___________________________", "Equipo"),
            ("Encargado: ___________________________", "Equipo"),
            ("Ausencias / cambios: ___________________________", "Equipo"),
        ]),
        ("Notas y promociones", [
            ("Producto destacado del día: ___________________________", "Tienda"),
            ("Promoción activa: ___________________________", "RRSS"),
            ("Eventos / visitas especiales: ___________________________", "Admin"),
            ("Incidencias del turno anterior: ___________________________", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-servicio.xlsx"))
    print("  ✓ BONUS-01-briefing-servicio.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS 02 — CALENDARIO ANUAL
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS 02 · Calendario Anual de Chocolatería", [
        "Cómo usar esta plantilla:",
        "▸ 15 fechas clave para chocolaterías con preparación especial.",
        "▸ Planifica producción, packaging y marketing con antelación.",
        "▸ Adapta las fechas a tu zona y mercado.",
    ])
    create_calendar_sheet(wb, "Calendario", ADMIN_COLOR, "Calendario Anual — Chocolatería", [
        ("Enero", "Rebajas post-Navidad, lanzar colección invierno, planificar San Valentín", "Trufas, bombones especiados, chocolate caliente", "Media"),
        ("Febrero", "San Valentín (14), producción corazones y edición limitada, entregas express", "Corazones, bombones premium, tabletas personalizadas", "Alta"),
        ("Marzo", "Día del Padre (19 España), preparación Pascua, talleres infantiles", "Figuras, huevos, monas de chocolate", "Alta"),
        ("Abril", "Semana Santa / Pascua, monas y huevos, colección primavera", "Huevos, conejos, bombones florales", "Alta"),
        ("Mayo", "Día de la Madre (1er domingo), edición regalo premium", "Bombones premium, cestas regalo, tabletas grabadas", "Alta"),
        ("Junio", "Inicio verano, adaptar catálogo (menos bombones, más tabletas)", "Tabletas con fruta, bombones en cámara", "Media"),
        ("Julio", "Temporada turística, packs regalo turista, colaboraciones locales", "Tabletas souvenir, bombones regionales", "Media"),
        ("Agosto", "Mantenimiento profundo obrador, vacaciones escalonadas, stock Navidad", "Producción reducida, planificación Q4", "Baja"),
        ("Septiembre", "Vuelta a la rutina, lanzar colección otoño, ferias gastronómicas", "Bombones de frutos secos, praliné, especiados", "Media"),
        ("Octubre", "Halloween (31), producción figuras terror, preparación Navidad intensa", "Figuras Halloween, bombones calabaza, tabletas oscuras", "Alta"),
        ("Noviembre", "Black Friday, preventa Navidad, producción de turrones y cestas", "Turrones, cestas corporativas, figuras navideñas", "Alta"),
        ("Diciembre", "NAVIDAD — máxima producción: turrones, figuras, cestas, regalos corporativos", "Turrones, figuras, bombones navideños, cestas", "Alta"),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-tareas.xlsx"))
    print("  ✓ BONUS-02-calendario-anual-tareas.xlsx")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\n🍫 Generando Kit de Tareas — Chocolatería / Obrador de Chocolate\n   → {OUTPUT_DIR}\n")
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_bonus_01()
    gen_bonus_02()
    print(f"\n✅ 9 archivos generados en {OUTPUT_DIR}\n")
