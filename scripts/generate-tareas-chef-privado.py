#!/usr/bin/env python3
"""
Generate Kit de Tareas Recurrentes — Chef Privado / Personal Chef (9 archivos).
AI Chef Pro — aichef.pro
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "dl", "kit-tareas-chef-privado"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── THEME COLORS ───────────────────────────────────────
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# Zone colors for Chef Privado
COCINA_COLOR = "FFF3E0"       # Orange light - kitchen/cooking
TRANSPORTE_COLOR = "E3F2FD"   # Blue light - transport/logistics
CLIENTE_COLOR = "F3E5F5"      # Purple light - client management
COMPRAS_COLOR = "E8F5E9"      # Green light - shopping/sourcing
SEGURIDAD_COLOR = "FFEBEE"    # Red light - food safety
SERVICIO_COLOR = "E0F7FA"     # Cyan light - service execution
ADMIN_COLOR = "FFF8E1"        # Amber light - admin/finance
MARKETING_COLOR = "F1F8E9"    # Light green - marketing
LIMPIEZA_COLOR = "EFEBE9"     # Brown light - cleaning
EQUIPO_COLOR = "E8EAF6"       # Indigo light - equipment

ZONE_COLORS = {
    "Cocina": COCINA_COLOR,
    "Prep": COCINA_COLOR,
    "Mise en place": COCINA_COLOR,
    "Producción": COCINA_COLOR,
    "Emplatado": COCINA_COLOR,
    "Transporte": TRANSPORTE_COLOR,
    "Vehículo": TRANSPORTE_COLOR,
    "Logística": TRANSPORTE_COLOR,
    "Carga": TRANSPORTE_COLOR,
    "Cliente": CLIENTE_COLOR,
    "Consulta": CLIENTE_COLOR,
    "Preferencias": CLIENTE_COLOR,
    "Seguimiento": CLIENTE_COLOR,
    "Compras": COMPRAS_COLOR,
    "Mercado": COMPRAS_COLOR,
    "Proveedores": COMPRAS_COLOR,
    "Stock": COMPRAS_COLOR,
    "Seguridad": SEGURIDAD_COLOR,
    "APPCC": SEGURIDAD_COLOR,
    "Temperatura": SEGURIDAD_COLOR,
    "Alergenos": SEGURIDAD_COLOR,
    "Servicio": SERVICIO_COLOR,
    "Mesa": SERVICIO_COLOR,
    "Evento": SERVICIO_COLOR,
    "Admin": ADMIN_COLOR,
    "Finanzas": ADMIN_COLOR,
    "Facturación": ADMIN_COLOR,
    "Agenda": ADMIN_COLOR,
    "Marketing": MARKETING_COLOR,
    "RRSS": MARKETING_COLOR,
    "Portfolio": MARKETING_COLOR,
    "Limpieza": LIMPIEZA_COLOR,
    "Post-servicio": LIMPIEZA_COLOR,
    "Equipo": EQUIPO_COLOR,
    "Kit": EQUIPO_COLOR,
    "General": LIGHT_GRAY,
}

# ─── FONTS & STYLES ────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
subtitle_font = Font(name="Calibri", size=11, color="888888", italic=True)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=GOLD)
data_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
small_font = Font(name="Calibri", size=9, color="999999")
checkbox_font = Font(name="Calibri", size=14)

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── SHARED FUNCTIONS ──────────────────────────────────
def add_instructions_sheet(wb, title, lines):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws["B2"].value = title
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=GOLD)
    for i, line in enumerate(lines, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("▸"):
            cell.font = Font(name="Calibri", size=11, color="555555")
        elif line == "":
            pass
        else:
            cell.font = Font(name="Calibri", size=12, bold=True, color="333333")


def create_task_sheet(wb, sheet_name, tab_color, title, sections):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Chef Privado / Personal Chef"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for section_title, tasks in sections:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = section_title
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["#", "Tarea", "Zona", "Responsable", "✓", "Hora", "Notas"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for idx, (task, zone) in enumerate(tasks, 1):
            zone_color = ZONE_COLORS.get(zone, LIGHT_GRAY)
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, fill_type="solid")

            ws.cell(row=row, column=1, value=idx).font = data_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=task).font = data_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=zone).font = data_font
            ws.cell(row=row, column=3).fill = zone_fill
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).fill = input_fill
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align

            check_cell = ws.cell(row=row, column=5)
            check_cell.font = checkbox_font
            check_cell.alignment = center_align
            check_cell.border = thin_border
            dv.add(check_cell)

            ws.cell(row=row, column=6).fill = input_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align

            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=7).alignment = left_align

            row += 1
        row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Firma responsable: _________________  Fecha: ___/___/______"
    ws[f"A{row}"].font = small_font
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


def create_blank_template_sheet(wb, sheet_name, tab_color, title, col_headers):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:G2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Chef Privado / Personal Chef"
    ws["A2"].font = subtitle_font

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    dv.prompt = "Marca ✓ si completada"
    ws.add_data_validation(dv)

    row = 4
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r in range(5, 30):
        for c in range(1, len(col_headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c == 1:
                cell.alignment = center_align
            elif c == 5:
                cell.font = checkbox_font
                cell.alignment = center_align
                dv.add(cell)
            elif c in (4, 6):
                cell.fill = input_fill
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.merge_cells(f"A31:G31")
    ws[f"A31"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A31"].font = small_font


def create_calendar_sheet(wb, sheet_name, tab_color, title, months_data):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:F2")
    ws["A2"].value = "AI Chef Pro · aichef.pro — Kit de Tareas: Chef Privado / Personal Chef"
    ws["A2"].font = subtitle_font

    row = 4
    cal_headers = ["Mes", "Acciones clave / Fechas señaladas", "Tipo de servicio", "Demanda", "✓", "Notas"]
    for col_idx, h in enumerate(cal_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Usa ✓, ✗ o —"
    ws.add_data_validation(dv)

    for month, actions, services, demand in months_data:
        ws.cell(row=row, column=1, value=month).font = bold_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=actions).font = data_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=services).font = data_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = thin_border

        demand_color = COCINA_COLOR if demand == "Alta" else TRANSPORTE_COLOR if demand == "Baja" else SERVICIO_COLOR
        ws.cell(row=row, column=4, value=demand).font = data_font
        ws.cell(row=row, column=4).fill = PatternFill(start_color=demand_color, end_color=demand_color, fill_type="solid")
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border

        check = ws.cell(row=row, column=5)
        check.font = checkbox_font
        check.alignment = center_align
        check.border = thin_border
        dv.add(check)

        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = left_align

        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "© 2026 AI Chef Pro · aichef.pro"
    ws[f"A{row}"].font = small_font


# ═══════════════════════════════════════════════════════════
# 01 — FICHA DE CLIENTE Y CONSULTA INICIAL
# ═══════════════════════════════════════════════════════════
def gen_01():
    wb = Workbook()
    add_instructions_sheet(wb, "01 · Ficha de Cliente y Consulta Inicial", [
        "Cómo usar esta plantilla:",
        "▸ Rellena una ficha por cada cliente nuevo.",
        "▸ Documenta alergias, preferencias e intolerencias ANTES del primer servicio.",
        "▸ Revisa la ficha antes de cada servicio con ese cliente.",
        "▸ Actualiza tras cada servicio con feedback recibido.",
        "",
        "Zonas de color:",
        "▸ Morado = Datos del cliente",
        "▸ Verde = Preferencias / dieta",
        "▸ Rojo = Alergias e intolerancias (CRÍTICO)",
        "▸ Ámbar = Logística (cocina, acceso, parking)",
    ])
    create_task_sheet(wb, "Ficha Cliente", "9C27B0", "Ficha de Cliente y Consulta Inicial", [
        ("📋 DATOS DEL CLIENTE", [
            ("Nombre completo del cliente", "Cliente"),
            ("Dirección completa del servicio", "Cliente"),
            ("Teléfono de contacto principal", "Cliente"),
            ("Email del cliente", "Cliente"),
            ("Instrucciones de acceso (portero, código, parking)", "Logística"),
            ("Número habitual de comensales", "Cliente"),
            ("Tipo de servicio solicitado (cena, meal prep, evento, clase)", "Cliente"),
            ("Presupuesto por servicio / por persona", "Cliente"),
            ("Frecuencia deseada (semanal, quincenal, evento puntual)", "Cliente"),
        ]),
        ("⚠️ ALERGIAS E INTOLERANCIAS (CRÍTICO)", [
            ("Registrar alergias del comensal 1 (14 alérgenos UE)", "Alergenos"),
            ("Registrar alergias del comensal 2", "Alergenos"),
            ("Registrar alergias del comensal 3", "Alergenos"),
            ("Registrar alergias del comensal 4", "Alergenos"),
            ("Intolerancias (lactosa, fructosa, histamina, etc.)", "Alergenos"),
            ("Restricciones religiosas o culturales", "Preferencias"),
            ("Dieta especial (vegano, keto, sin gluten, bajo en sal, FODMAP)", "Preferencias"),
            ("Firma del cliente confirmando alergias declaradas", "Alergenos"),
        ]),
        ("🍽️ PREFERENCIAS CULINARIAS", [
            ("Cocinas favoritas (mediterránea, japonesa, peruana, fusión...)", "Preferencias"),
            ("Ingredientes que NO les gustan", "Preferencias"),
            ("Ingredientes favoritos / caprichos", "Preferencias"),
            ("Nivel de picante tolerado", "Preferencias"),
            ("Preferencias de presentación (casual, fine dining, familiar)", "Preferencias"),
            ("Vinos / maridaje (incluir o no, preferencias)", "Preferencias"),
            ("Postres (sí/no, preferencias)", "Preferencias"),
        ]),
        ("🏠 EQUIPAMIENTO DE LA COCINA DEL CLIENTE", [
            ("Tipo de cocina (gas, inducción, vitro, mixta)", "Logística"),
            ("Horno disponible (sí/no, tipo)", "Logística"),
            ("Batidora / robot de cocina disponible", "Logística"),
            ("Espacio de trabajo disponible (amplio, limitado)", "Logística"),
            ("Frigorífico / congelador (capacidad)", "Logística"),
            ("Vajilla y cristalería (servicio propio o del cliente)", "Logística"),
            ("Acceso a agua caliente y lavavajillas", "Logística"),
            ("Notas sobre limitaciones de la cocina", "Logística"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "01-ficha-cliente-consulta.xlsx"))
    print("✅ 01-ficha-cliente-consulta.xlsx")


# ═══════════════════════════════════════════════════════════
# 02 — PLANIFICACIÓN DE MENÚ Y COMPRAS
# ═══════════════════════════════════════════════════════════
def gen_02():
    wb = Workbook()
    add_instructions_sheet(wb, "02 · Planificación de Menú y Compras", [
        "Cómo usar esta plantilla:",
        "▸ Planifica el menú completo antes de cada servicio.",
        "▸ Calcula raciones y genera la lista de compras.",
        "▸ Controla el coste por persona para mantener márgenes.",
        "▸ Envía la propuesta de menú al cliente para aprobación.",
        "",
        "Consejo profesional:",
        "▸ Planifica con 48-72h de antelación para producto fresco.",
        "▸ Ten siempre una lista B de platos alternativos.",
    ])
    create_task_sheet(wb, "Planificación Menú", "4CAF50", "Planificación de Menú por Servicio", [
        ("📝 PROPUESTA DE MENÚ", [
            ("Definir tipo de servicio (cena, meal prep, evento, brunch)", "Cocina"),
            ("Número de comensales confirmados", "Cliente"),
            ("Revisar ficha de alergias/preferencias del cliente", "Alergenos"),
            ("Diseñar aperitivo / snack de bienvenida", "Cocina"),
            ("Diseñar entrante 1", "Cocina"),
            ("Diseñar entrante 2 (si aplica)", "Cocina"),
            ("Diseñar plato principal", "Cocina"),
            ("Diseñar guarniciones", "Cocina"),
            ("Diseñar postre", "Cocina"),
            ("Seleccionar maridaje / bebidas (si incluido)", "Cocina"),
            ("Verificar que NINGÚN plato contiene alérgenos del cliente", "Alergenos"),
            ("Enviar propuesta al cliente para aprobación", "Cliente"),
            ("Confirmar menú definitivo con el cliente", "Cliente"),
        ]),
        ("🛒 LISTA DE COMPRAS", [
            ("Calcular raciones por comensal (con margen 10-15%)", "Compras"),
            ("Listar ingredientes del entrante", "Compras"),
            ("Listar ingredientes del principal", "Compras"),
            ("Listar ingredientes del postre", "Compras"),
            ("Listar ingredientes de salsas/bases/fondos", "Compras"),
            ("Verificar stock de despensa portátil (aceite, sal, especias)", "Stock"),
            ("Identificar ingredientes que requieren pedido a proveedor", "Proveedores"),
            ("Comparar precios mercado vs proveedor", "Compras"),
            ("Comprar producto fresco (mercado, pescadería, carnicería)", "Mercado"),
            ("Registrar gasto total de compras para este servicio", "Compras"),
        ]),
        ("💰 CONTROL DE COSTE", [
            ("Calcular coste total de ingredientes", "Finanzas"),
            ("Calcular coste por comensal", "Finanzas"),
            ("Calcular margen bruto del servicio", "Finanzas"),
            ("Verificar que el margen es ≥60% del precio cobrado", "Finanzas"),
            ("Registrar gastos de transporte", "Finanzas"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "02-planificacion-menu-compras.xlsx"))
    print("✅ 02-planificacion-menu-compras.xlsx")


# ═══════════════════════════════════════════════════════════
# 03 — CHECKLIST DE EQUIPO Y TRANSPORTE
# ═══════════════════════════════════════════════════════════
def gen_03():
    wb = Workbook()
    add_instructions_sheet(wb, "03 · Checklist de Equipo y Transporte", [
        "Cómo usar esta plantilla:",
        "▸ Revisa ANTES de salir de casa/base para cada servicio.",
        "▸ Marca cada ítem al cargarlo en el vehículo.",
        "▸ Usa la misma lista al recoger para no dejar nada en casa del cliente.",
        "",
        "Regla de oro:",
        "▸ Si no está en la lista, no va en el coche.",
        "▸ Revisa nevera portátil 30 min antes de salir.",
    ])
    create_task_sheet(wb, "Equipo Transporte", "2196F3", "Checklist de Equipo y Transporte", [
        ("🔪 KIT DE CUCHILLOS Y UTENSILIOS", [
            ("Estuche de cuchillos completo (chef, puntilla, fileteador, sierra)", "Kit"),
            ("Tabla de corte profesional", "Kit"),
            ("Pelador, mandolina, microplane", "Kit"),
            ("Espátulas, cucharas, pinzas de emplatar", "Kit"),
            ("Batidora de mano / túrmix", "Kit"),
            ("Termómetro de cocina (sonda)", "Kit"),
            ("Soplete de cocina", "Kit"),
            ("Sifón (si necesario para el menú)", "Kit"),
        ]),
        ("🍳 SARTENES Y OLLAS", [
            ("Sartén antiadherente 24cm + 28cm", "Kit"),
            ("Sartén de hierro / parrilla (si necesaria)", "Kit"),
            ("Olla / cacerola mediana", "Kit"),
            ("Cazo para salsas", "Kit"),
            ("Bandeja de horno (si necesaria)", "Kit"),
        ]),
        ("❄️ TRANSPORTE FRÍO", [
            ("Nevera portátil con acumuladores de frío", "Transporte"),
            ("Bolsas isotérmicas para producto fresco", "Transporte"),
            ("Verificar temperatura nevera portátil (<5°C)", "Temperatura"),
            ("Hielo extra para emergencias", "Transporte"),
            ("Separar crudos de cocinados en transporte", "Seguridad"),
        ]),
        ("🧴 MINI-DESPENSA PORTÁTIL", [
            ("Aceite de oliva virgen extra", "Stock"),
            ("Sal Maldon + sal fina", "Stock"),
            ("Pimienta (molinillo)", "Stock"),
            ("Vinagre (jerez, balsámico, de arroz)", "Stock"),
            ("Limones y limas", "Stock"),
            ("Mantequilla", "Stock"),
            ("Especias básicas (pimentón, comino, cúrcuma, tomillo)", "Stock"),
            ("Azúcar, miel", "Stock"),
        ]),
        ("👔 TEXTILES Y PRESENTACIÓN", [
            ("Delantal limpio (2 unidades)", "Kit"),
            ("Paños de cocina (mínimo 4)", "Kit"),
            ("Guantes desechables (caja)", "Kit"),
            ("Gorro/cofia de cocina", "Kit"),
            ("Servilletas de tela (si servicio fine dining)", "Kit"),
        ]),
        ("🚗 VERIFICACIÓN DEL VEHÍCULO", [
            ("Verificar combustible suficiente", "Vehículo"),
            ("GPS configurado con dirección del cliente", "Vehículo"),
            ("Verificar hora de llegada (30-60 min antes del servicio)", "Vehículo"),
            ("Parking confirmado en dirección del cliente", "Logística"),
            ("Carga completa verificada (nada en casa)", "Carga"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "03-equipo-transporte.xlsx"))
    print("✅ 03-equipo-transporte.xlsx")


# ═══════════════════════════════════════════════════════════
# 04 — SEGURIDAD ALIMENTARIA / APPCC MÓVIL
# ═══════════════════════════════════════════════════════════
def gen_04():
    wb = Workbook()
    add_instructions_sheet(wb, "04 · Seguridad Alimentaria / APPCC Móvil", [
        "Cómo usar esta plantilla:",
        "▸ Registra temperaturas en CADA servicio.",
        "▸ Verifica alérgenos ANTES de cada plato.",
        "▸ Documenta la trazabilidad de ingredientes.",
        "▸ Guarda estos registros 2 años (requisito legal UE).",
        "",
        "IMPORTANTE:",
        "▸ Como chef privado, TÚ eres responsable de la seguridad alimentaria.",
        "▸ Un incidente con alérgenos puede acabar con tu negocio.",
    ])
    create_task_sheet(wb, "APPCC Móvil", "F44336", "Seguridad Alimentaria — APPCC Móvil", [
        ("🌡️ CONTROL DE TEMPERATURAS", [
            ("Registrar temperatura nevera portátil al salir (<5°C)", "Temperatura"),
            ("Registrar temperatura producto fresco al llegar al cliente", "Temperatura"),
            ("Verificar temperatura frigorífico del cliente", "Temperatura"),
            ("Control de temperatura durante mise en place", "Temperatura"),
            ("Verificar temperatura interna de carnes/pescados cocinados (>75°C)", "Temperatura"),
            ("Control de temperatura de producto en espera (<5°C o >65°C)", "Temperatura"),
            ("Registrar hora inicio y fin de zona de peligro (5-65°C)", "Temperatura"),
        ]),
        ("⚠️ GESTIÓN DE ALÉRGENOS", [
            ("Revisar ficha de alergias del cliente ANTES de cocinar", "Alergenos"),
            ("Verificar etiquetado de todos los ingredientes comprados", "Alergenos"),
            ("Separar utensilios para preparación sin alérgenos", "Alergenos"),
            ("Evitar contaminación cruzada (tablas, cuchillos, aceites)", "Alergenos"),
            ("Confirmar con el cliente: 'El menú de hoy no contiene [X]'", "Alergenos"),
            ("Si meal prep: etiquetar cada envase con alérgenos presentes", "Alergenos"),
        ]),
        ("🧼 HIGIENE Y LIMPIEZA", [
            ("Lavado de manos al llegar a la cocina del cliente", "Seguridad"),
            ("Uniforme limpio y completo (delantal, gorro)", "Seguridad"),
            ("Desinfectar superficies de trabajo en cocina ajena", "Limpieza"),
            ("Verificar estado de limpieza de utensilios del cliente (si se usan)", "Limpieza"),
            ("Separar residuos (orgánico, envases, vidrio)", "Limpieza"),
            ("Desinfectar zona de trabajo al terminar", "Limpieza"),
        ]),
        ("📦 TRAZABILIDAD Y ETIQUETADO", [
            ("Conservar tickets de compra de ingredientes", "Seguridad"),
            ("Registrar proveedor y lote de productos clave (carne, pescado)", "Seguridad"),
            ("Si meal prep: etiquetar con fecha elaboración y caducidad", "Seguridad"),
            ("Si meal prep: incluir instrucciones de conservación y recalentado", "Seguridad"),
            ("Registrar incidencias de seguridad alimentaria (si las hubiera)", "Seguridad"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "04-seguridad-alimentaria-appcc.xlsx"))
    print("✅ 04-seguridad-alimentaria-appcc.xlsx")


# ═══════════════════════════════════════════════════════════
# 05 — CHECKLIST DE SERVICIO (PRE / DURANTE / POST)
# ═══════════════════════════════════════════════════════════
def gen_05():
    wb = Workbook()
    add_instructions_sheet(wb, "05 · Checklist de Servicio (Pre / Durante / Post)", [
        "Cómo usar esta plantilla:",
        "▸ Sigue este checklist en CADA servicio.",
        "▸ Pre-servicio: desde que llegas hasta que empiezas a cocinar.",
        "▸ Durante: timing de platos, presentación, servicio.",
        "▸ Post-servicio: limpieza, recogida, cierre.",
        "",
        "Clave de éxito:",
        "▸ Llega siempre 45-60 min antes de la hora de servicio.",
        "▸ Deja la cocina MEJOR de como la encontraste.",
    ])
    create_task_sheet(wb, "Pre-Servicio", "00BCD4", "Checklist Pre-Servicio", [
        ("🏠 LLEGADA Y MONTAJE", [
            ("Llegar 45-60 min antes del servicio", "Logística"),
            ("Presentarse al cliente, confirmar hora de servicio", "Cliente"),
            ("Inspeccionar cocina del cliente (estado, equipo disponible)", "Cocina"),
            ("Descargar equipo y producto del vehículo", "Carga"),
            ("Verificar temperatura del producto transportado", "Temperatura"),
            ("Montar estación de trabajo (utensilios, tabla, cuchillos)", "Mise en place"),
            ("Organizar ingredientes por orden de uso", "Mise en place"),
            ("Verificar funcionamiento de fogones/horno del cliente", "Cocina"),
        ]),
        ("👨‍🍳 MISE EN PLACE", [
            ("Lavar y preparar verduras y hierbas", "Prep"),
            ("Porcionar proteínas (carne, pescado)", "Prep"),
            ("Preparar bases, salsas, reducciones", "Prep"),
            ("Preparar guarniciones", "Prep"),
            ("Preparar elementos del postre", "Prep"),
            ("Verificar que todo está listo para el timing del servicio", "Prep"),
            ("Confirmar hora exacta de primer pase con el cliente", "Cliente"),
        ]),
    ])
    create_task_sheet(wb, "Durante Servicio", "FF9800", "Checklist Durante el Servicio", [
        ("🍽️ EJECUCIÓN Y SERVICIO", [
            ("Servir aperitivo / snack de bienvenida", "Servicio"),
            ("Primer pase: entrante 1 (emplatar, presentar)", "Emplatado"),
            ("Segundo pase: entrante 2 (timing correcto)", "Emplatado"),
            ("Tercer pase: plato principal + guarniciones", "Emplatado"),
            ("Limpiar cocina entre pases (no acumular suciedad)", "Limpieza"),
            ("Servir postre", "Emplatado"),
            ("Ofrecer café / infusión (si incluido)", "Servicio"),
            ("Verificar satisfacción del cliente durante el servicio", "Cliente"),
            ("Gestionar imprevistos (ingrediente faltante, cambio de timing)", "Servicio"),
        ]),
    ])
    create_task_sheet(wb, "Post-Servicio", "795548", "Checklist Post-Servicio", [
        ("🧹 LIMPIEZA Y RECOGIDA", [
            ("Limpiar TODA la cocina del cliente (fogones, superficies, fregadero)", "Limpieza"),
            ("Lavar y secar toda la vajilla utilizada", "Limpieza"),
            ("Limpiar horno/microondas si se usó", "Limpieza"),
            ("Barrer/fregar suelo de cocina", "Limpieza"),
            ("Sacar basura y residuos (separados)", "Limpieza"),
            ("Almacenar sobras con instrucciones para el cliente", "Post-servicio"),
            ("Dejar la cocina MEJOR de como la encontraste", "Limpieza"),
        ]),
        ("📦 RECOGIDA DE EQUIPO", [
            ("Inventario de equipo: verificar que no dejas NADA", "Kit"),
            ("Limpiar y guardar cuchillos en estuche", "Kit"),
            ("Limpiar y empacar sartenes/ollas", "Kit"),
            ("Recoger despensa portátil sobrante", "Stock"),
            ("Cargar todo en el vehículo", "Carga"),
            ("Despedirse del cliente y confirmar satisfacción", "Cliente"),
        ]),
        ("📸 DOCUMENTACIÓN", [
            ("Fotografiar platos para portfolio/redes (si permitido)", "Portfolio"),
            ("Anotar feedback del cliente", "Seguimiento"),
            ("Registrar incidencias o notas para próximo servicio", "Seguimiento"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "05-checklist-servicio.xlsx"))
    print("✅ 05-checklist-servicio.xlsx")


# ═══════════════════════════════════════════════════════════
# 06 — SEGUIMIENTO Y FIDELIZACIÓN DE CLIENTES
# ═══════════════════════════════════════════════════════════
def gen_06():
    wb = Workbook()
    add_instructions_sheet(wb, "06 · Seguimiento y Fidelización de Clientes", [
        "Cómo usar esta plantilla:",
        "▸ Completa después de cada servicio.",
        "▸ Envía agradecimiento en las primeras 24 horas.",
        "▸ Solicita reseña a clientes satisfechos.",
        "▸ Usa el historial para personalizar futuros servicios.",
        "",
        "El chef privado vive de la recurrencia:",
        "▸ Un cliente contento repite y te recomienda.",
        "▸ El seguimiento profesional te diferencia del 90% de la competencia.",
    ])
    create_task_sheet(wb, "Post-Servicio", "9C27B0", "Seguimiento Post-Servicio", [
        ("📱 PRIMERAS 24 HORAS", [
            ("Enviar mensaje de agradecimiento al cliente (WhatsApp/email)", "Seguimiento"),
            ("Preguntar si disfrutaron de la experiencia", "Seguimiento"),
            ("Enviar fotos de los platos al cliente (si las tiene)", "Seguimiento"),
            ("Solicitar reseña en Google / redes sociales", "Seguimiento"),
            ("Registrar feedback recibido en ficha del cliente", "Seguimiento"),
        ]),
        ("📅 FIDELIZACIÓN (SEMANA SIGUIENTE)", [
            ("Proponer fecha para próximo servicio", "Seguimiento"),
            ("Enviar propuesta de menú para siguiente ocasión", "Seguimiento"),
            ("Ofrecer descuento por recurrencia (si aplica)", "Seguimiento"),
            ("Actualizar ficha de preferencias con nuevos datos", "Preferencias"),
            ("Añadir al calendario de fechas especiales del cliente", "Agenda"),
        ]),
        ("🎂 FECHAS ESPECIALES", [
            ("Registrar cumpleaños del cliente y familiares", "Cliente"),
            ("Registrar aniversarios importantes", "Cliente"),
            ("Enviar felicitación + oferta especial en fechas señaladas", "Seguimiento"),
            ("Proponer menú especial para celebraciones", "Seguimiento"),
        ]),
    ])
    create_task_sheet(wb, "Historial Servicios", "FF9800", "Historial de Servicios por Cliente", [
        ("📊 REGISTRO DE SERVICIOS", [
            ("Fecha del servicio", "Admin"),
            ("Tipo de servicio (cena, meal prep, evento, clase)", "Admin"),
            ("Número de comensales", "Admin"),
            ("Menú servido (resumen)", "Admin"),
            ("Importe facturado", "Finanzas"),
            ("Coste de ingredientes", "Finanzas"),
            ("Margen del servicio", "Finanzas"),
            ("Satisfacción del cliente (1-5)", "Seguimiento"),
            ("Notas / feedback / incidencias", "Seguimiento"),
            ("Repetición: ¿ha reservado siguiente servicio?", "Seguimiento"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "06-seguimiento-fidelizacion.xlsx"))
    print("✅ 06-seguimiento-fidelizacion.xlsx")


# ═══════════════════════════════════════════════════════════
# 07 — ADMINISTRACIÓN DEL AUTÓNOMO
# ═══════════════════════════════════════════════════════════
def gen_07():
    wb = Workbook()
    add_instructions_sheet(wb, "07 · Administración del Autónomo", [
        "Cómo usar esta plantilla:",
        "▸ Lleva el control de facturación, gastos e impuestos.",
        "▸ Actualiza semanalmente para no acumular trabajo.",
        "▸ Revisa mensualmente los márgenes por cliente.",
        "",
        "Fechas clave del autónomo en España:",
        "▸ Trimestral: modelo 303 (IVA), modelo 130 (IRPF).",
        "▸ Anual: modelo 390, declaración de la renta.",
        "▸ Seguro de responsabilidad civil: OBLIGATORIO.",
    ])
    create_task_sheet(wb, "Facturación", "FF9800", "Control de Facturación", [
        ("💶 FACTURACIÓN MENSUAL", [
            ("Emitir factura de cada servicio (en 48h)", "Facturación"),
            ("Incluir en factura: servicio + ingredientes + transporte", "Facturación"),
            ("Verificar IVA correcto (10% servicios alimentación / 21% otros)", "Facturación"),
            ("Enviar factura al cliente por email", "Facturación"),
            ("Verificar cobros recibidos vs facturas emitidas", "Facturación"),
            ("Registrar cobros pendientes y hacer seguimiento", "Facturación"),
        ]),
        ("📊 CONTROL DE GASTOS", [
            ("Registrar gastos de ingredientes por servicio", "Finanzas"),
            ("Registrar gastos de transporte (gasolina, peajes, parking)", "Finanzas"),
            ("Registrar desgaste de equipo y reposición", "Finanzas"),
            ("Registrar cuota de autónomo mensual", "Finanzas"),
            ("Registrar seguro de responsabilidad civil", "Finanzas"),
            ("Registrar gastos de marketing y web", "Finanzas"),
            ("Separar gastos deducibles de no deducibles", "Finanzas"),
        ]),
        ("📅 OBLIGACIONES FISCALES", [
            ("Modelo 303 — IVA trimestral (ene, abr, jul, oct)", "Admin"),
            ("Modelo 130 — IRPF trimestral", "Admin"),
            ("Modelo 390 — Resumen anual IVA (enero)", "Admin"),
            ("Declaración de la renta (abril-junio)", "Admin"),
            ("Renovar seguro de responsabilidad civil", "Admin"),
            ("Renovar carnet de manipulador de alimentos", "Admin"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "07-administracion-autonomo.xlsx"))
    print("✅ 07-administracion-autonomo.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS 01 — BRIEFING PRE-SERVICIO
# ═══════════════════════════════════════════════════════════
def gen_bonus_01():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS · Briefing Pre-Servicio", [
        "Cómo usar esta plantilla:",
        "▸ Rellena 24h antes de cada servicio.",
        "▸ Es tu 'hoja de ruta' para el día.",
        "▸ Revísalo antes de salir de casa.",
        "▸ Si trabajas con ayudante, comparte este briefing.",
        "",
        "Un buen briefing evita:",
        "▸ Olvidar ingredientes o equipo.",
        "▸ Llegar tarde por mala planificación.",
        "▸ Servir algo con un alérgeno no controlado.",
    ])
    create_task_sheet(wb, "Briefing", "E91E63", "Briefing Pre-Servicio", [
        ("📋 INFORMACIÓN DEL SERVICIO", [
            ("Cliente: [nombre]", "Cliente"),
            ("Dirección: [dirección completa]", "Logística"),
            ("Hora de llegada: [hora]", "Logística"),
            ("Hora de servicio: [hora]", "Servicio"),
            ("Nº comensales: [número]", "Cliente"),
            ("Tipo: cena / meal prep / evento / clase", "Servicio"),
        ]),
        ("⚠️ ALERTAS CRÍTICAS", [
            ("ALERGIAS: [listar todas]", "Alergenos"),
            ("INTOLERANCIAS: [listar todas]", "Alergenos"),
            ("DIETA ESPECIAL: [vegano/keto/sin gluten/etc.]", "Preferencias"),
            ("INGREDIENTES PROHIBIDOS: [listar]", "Alergenos"),
        ]),
        ("🍽️ MENÚ DEL DÍA", [
            ("Aperitivo: [descripción]", "Cocina"),
            ("Entrante 1: [descripción]", "Cocina"),
            ("Entrante 2: [descripción]", "Cocina"),
            ("Principal: [descripción]", "Cocina"),
            ("Postre: [descripción]", "Cocina"),
            ("Bebidas / maridaje: [descripción]", "Cocina"),
        ]),
        ("✅ VERIFICACIÓN FINAL", [
            ("Compras completadas", "Compras"),
            ("Equipo cargado en vehículo", "Carga"),
            ("Nevera portátil a temperatura", "Temperatura"),
            ("GPS configurado", "Vehículo"),
            ("Fotos del menú anterior del cliente revisadas", "Seguimiento"),
        ]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-01-briefing-pre-servicio.xlsx"))
    print("✅ BONUS-01-briefing-pre-servicio.xlsx")


# ═══════════════════════════════════════════════════════════
# BONUS 02 — CALENDARIO ANUAL DE DEMANDA
# ═══════════════════════════════════════════════════════════
def gen_bonus_02():
    wb = Workbook()
    add_instructions_sheet(wb, "BONUS · Calendario Anual de Demanda", [
        "Cómo usar esta plantilla:",
        "▸ Planifica tu año de trabajo con las temporadas de demanda.",
        "▸ Anticipa los picos: festivos, verano (costas), Navidad.",
        "▸ Usa los meses bajos para marketing, formación y descanso.",
        "",
        "Dato clave:",
        "▸ El 40% de la facturación de un chef privado se concentra en",
        "  Navidad, Nochevieja, San Valentín, Semana Santa y verano.",
    ])
    create_calendar_sheet(wb, "Calendario Anual", "E91E63", "Calendario Anual de Demanda — Chef Privado", [
        ("Enero", "Detox post-navidad, menús saludables. Resolver compromisos Año Nuevo. Planificar año.", "Meal prep detox, cenas íntimas", "Media"),
        ("Febrero", "San Valentín (14 feb) = PICO. Cenas románticas premium. Preparar temporada Carnaval.", "Cenas románticas, maridaje", "Alta"),
        ("Marzo-Abril", "Semana Santa = PICO en zonas costeras. Villas, yates. Comuniones empiezan.", "Eventos familiares, villas", "Alta"),
        ("Mayo", "Día de la Madre. Comuniones. Eventos corporativos de primavera.", "Brunch, eventos, comuniones", "Alta"),
        ("Junio", "Bodas. Graduaciones. Cenas de empresa fin de temporada. Inicio temporada yates.", "Eventos, yates, bodas", "Alta"),
        ("Julio", "TEMPORADA ALTA en costa (Ibiza, Mallorca, Marbella, Costa del Sol). Villas y yates.", "Villas, yates, fiestas privadas", "Alta"),
        ("Agosto", "Pico máximo en costa. Fiestas privadas, barbacoas premium, pop-ups.", "Villas, yates, barbacoas", "Alta"),
        ("Septiembre", "Vuelta al cole. Clientes corporativos reactivan. Meal prep post-verano.", "Meal prep, corporativo", "Media"),
        ("Octubre", "Halloween. Cenas temáticas. Eventos de otoño. Temporada de caza.", "Cenas temáticas, caza, setas", "Media"),
        ("Noviembre", "Black Friday para vender bonos regalo. Preparar temporada navideña.", "Marketing, bonos regalo", "Baja"),
        ("Diciembre", "PICO MÁXIMO. Cenas de empresa, Nochebuena, Nochevieja. Máxima capacidad.", "Cenas empresa, Navidad, NYE", "Alta"),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "BONUS-02-calendario-anual-demanda.xlsx"))
    print("✅ BONUS-02-calendario-anual-demanda.xlsx")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\n📁 Generando en: {OUTPUT_DIR}\n")
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_bonus_01()
    gen_bonus_02()
    print(f"\n✅ 9 archivos generados en {OUTPUT_DIR}")
