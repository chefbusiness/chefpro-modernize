#!/usr/bin/env python3
"""
grupo_b.py — §3 de `kit-escandallos-v2-SPEC.md`: lo específico de 04, 05, 06 y 08.

  · 04 cócteles (DOM-06/COM-23/DOM-22/DOM-24/COM-33/TEC-30/COM-29): las
    cantidades de destilado ya vienen en cl del motor (§1.1 + PARCHES) y el
    Factor L→cl las convierte; aquí se añade la pestaña «Formatos de Compra»
    —el puente entre el precio de la BOTELLA que trae la factura (70 cl los
    destilados, 75 cl el vino) y el €/L que pide la columna «Precio/Ud»—, se
    renombran los rótulos de «plato» a «cóctel», el coste de elaboración pasa
    a ser «Merma y hielo (%)» al 5 % y el rango de food cost de barra queda en
    20-25 % en un único sitio.
  · 05 pastelería (DOM-04/TEC-05/COM-01/TEC-20/COM-13): «Nº de raciones» pasa a
    «Rendimiento (uds por tanda)» con el valor del título (12/20/30) y el
    bloque de resultado a «COSTE POR UNIDAD» / «PVP POR UNIDAD»; mermas de
    obrador reales (12 % cobertura, 8 % harinas). El colorante de 2 g ya lo
    corrige el motor (PARCHES: 0,04 botes de 50 g).
  · 06 catering (DOM-12/TEC-08/COM-05/DOM-19/TEC-21/TEC-07/COM-19/TEC-24/
    COM-16): la pestaña pasa a llamarse «Cocktail (por persona)» con aviso en
    A2, el Presupuesto se reconstruye en la convención del kit (PVP = coste /
    food cost objetivo) con el personal y el menaje parametrizados en celdas
    verdes, y se añaden las hojas «Checklist Evento» (desplegable ✓/—/N/A +
    contador) y «Presupuesto Cliente» (sin costes ni margen).
  · 08 food truck (DOM-07/TEC-06/COM-03): hoja «Punto de Equilibrio» con los
    costes fijos del día, el margen de contribución medio ponderado leído de
    las tres pestañas de escandallo y las unidades/día necesarias.

Casi todo va en `post()`, DESPUÉS del motor: las hojas nuevas referencian el
bloque de resultado (`motor.bloque`), que sólo existe con el layout final.

La ÚNICA excepción es el renombrado de «Cocktail 50 pax», que va en `pre()`.
Tiene que ser antes del motor porque `motor.REGISTRO` anota cada fórmula con el
título que la hoja tenía al escribirla, y `main.verificar_cache` la busca luego
por ese título en el fichero guardado: renombrar después dejaría las ~90
fórmulas del escandallo del 06 apuntando a una hoja que ya no existe («hoja
ausente») y el gate en rojo. El efecto secundario —que `remapear_referencias`
deje de reconocer la referencia cruzada de `Presupuesto`!C6— es inocuo: esta
misma pasada reescribe la hoja «Presupuesto» entera.

IDEMPOTENTE: las hojas nuevas se vacían y se vuelven a escribir enteras desde
cero en cada pasada, y los rótulos renombrados los reconoce `motor._canon`
(ALIAS_ROTULOS), así que el motor de la 2.ª pasada recupera de ellos la misma
plantilla de estilo y el mismo valor del cliente.
"""
import copy

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import motor

FICHEROS = ['04-cocktails-bebidas.xlsx', '05-pasteleria.xlsx',
            '06-catering.xlsx', '08-food-truck.xlsx']

DV_MARCA = '"✓,—,N/A"'
MARCA_COL = '✓ Completada'
HOLGURA = 4          # filas libres del checklist, dentro del COUNTIF

FMT_H = '0.0'        # horas
FMT_ENT = motor.FMT_ENT

HOJA_06_VIEJA = 'Cocktail 50 pax'
HOJA_06 = 'Cocktail (por persona)'


# ==========================================================================
# 04 · Formatos de compra (DOM-22)
# ==========================================================================
# Producto, formato de compra (cl), precio de la botella (€), dónde se usa.
# El €/L se calcula en la hoja y coincide con el «Precio/Ud» que ya trae cada
# escandallo: 19,60 € la botella de 70 cl SON los 28 €/L de Gin Tonic!D5.
BOTELLAS = [
    ('Ginebra premium', 70, 19.60, 'Gin Tonic Premium'),
    ('Ron blanco', 70, 10.50, 'Mojito Clásico'),
    ('Tequila reposado', 70, 15.40, 'Margarita'),
    ('Triple Sec / Cointreau', 70, 12.60, 'Margarita'),
    ('Aperol', 70, 9.80, 'Aperol Spritz'),
    ('Prosecco', 75, 6.00, 'Aperol Spritz'),
    ('Vino blanco de copeo', 75, 4.50, '(añade tus referencias debajo)'),
]
BOTELLAS_LIBRES = 5
# TEC-R2-15: la hoja calculaba el €/L correcto pero NADIE lo leía — había que
# copiarlo a mano a la columna «Precio/Ud» del cóctel. Hoy los números
# coinciden porque se cargaron a la vez; en cuanto el cliente actualizase el
# precio de una botella, los escandallos seguirían costando al precio viejo sin
# un solo aviso, que es exactamente la desviación silenciosa que el kit vende
# que evita. Índice de BOTELLAS → (pestaña, fila) del escandallo que lo usa.
ENLACES_PRECIO = {
    0: ('Gin Tonic Premium', 5),      # ginebra
    1: ('Mojito Clásico', 5),         # ron blanco
    2: ('Margarita', 5),              # tequila
    3: ('Margarita', 6),              # triple sec
    4: ('Aperol Spritz', 5),          # aperol
    5: ('Aperol Spritz', 6),          # prosecco
}
AZUL_ENLACE = 'DDEBF7'   # celda VINCULADA: ni verde (editable) ni blanca

MERMA_HIELO = 0.05   # overpour sin jigger + dilución (DOM-22)

# ==========================================================================
# 05 · Rendimiento por tanda (DOM-04/TEC-05/COM-01) y mermas de obrador
# ==========================================================================
RENDIMIENTO_05 = {
    'Tarta Chocolate': 12,
    'Croissants': 20,
    'Macarons': 30,
}
# hoja → {fila: merma}. Filas PRECARGADAS: el motor respeta su valor escrito,
# así que la merma «real de pastelería» que publica la landing hay que
# escribirla aquí (TEC-20/COM-13). Columna H = Merma en el layout v2.
MERMAS_05 = {
    'Tarta Chocolate': {5: 0.12,    # chocolate negro 70 % — pérdida de temperado
                        9: 0.08},   # harina floja — evaporación de horneado
    'Croissants': {5: 0.08},        # harina de fuerza
    'Macarons': {10: 0.12},         # chocolate blanco de la ganache
}

# ==========================================================================
# 06 · Checklist de evento (TEC-24/COM-16)
# ==========================================================================
CHK_HDR = ['#', 'Tarea', 'Bloque', 'Responsable', 'Cuándo', MARCA_COL]
CHK_ANCHOS = [5, 58, 16, 18, 14, 14]
CHECKLIST = [
    ('Confirmar por escrito el nº definitivo de comensales', 'Timings', 'Comercial', 'D-7'),
    ('Cerrar menú definitivo y recoger alérgenos e intolerancias', 'Timings', 'Jefe de cocina', 'D-7'),
    ('Cursar pedido a proveedores con fecha y hora de entrega', 'Timings', 'Compras', 'D-5'),
    ('Confirmar hora de acceso al espacio y muelle de carga', 'Timings', 'Producción', 'D-3'),
    ('Producir las elaboraciones que aguantan 24 h', 'Timings', 'Jefe de cocina', 'D-1'),
    ('Cargar el vehículo siguiendo la hoja de ruta', 'Timings', 'Producción', 'D · mañana'),
    ('Montaje en el espacio (buffet, barras, office)', 'Timings', 'Jefe de sala', 'D · H-3 h'),
    ('Briefing con el equipo de sala y reparto de rangos', 'Timings', 'Jefe de sala', 'D · H-1 h'),
    ('Servicio', 'Timings', 'Todo el equipo', 'D · hora H'),
    ('Desmontaje, recuento de material y salida', 'Timings', 'Jefe de sala', 'D · H+1 h'),

    ('Nº de camareros según el Presupuesto (1 por cada 22 pax)', 'Personal', 'Jefe de sala', 'D-3'),
    ('Jefe de sala asignado, con el teléfono del cliente', 'Personal', 'Comercial', 'D-3'),
    ('Cocina in situ: cuántos cocineros y en qué horario', 'Personal', 'Jefe de cocina', 'D-3'),
    ('Uniformidad completa y una muda de repuesto', 'Personal', 'Jefe de sala', 'D-1'),
    ('Partes de horas firmados al terminar el servicio', 'Personal', 'Jefe de sala', 'D · H+1 h'),

    ('Vajilla, cristalería y cubertería contadas y embaladas', 'Menaje', 'Producción', 'D-1'),
    ('Mantelería, faldones y decoración del mobiliario', 'Menaje', 'Producción', 'D-1'),
    ('Bandejas, pinzas y utillaje de pase', 'Menaje', 'Producción', 'D-1'),
    ('Desechables: servilletas, palillos y bolsas de residuos', 'Menaje', 'Producción', 'D-1'),
    ('Menaje de repuesto: un 10 % sobre el nº de comensales', 'Menaje', 'Producción', 'D-1'),

    ('Vehículo isotermo reservado, revisado y con combustible', 'Transporte', 'Logística', 'D-2'),
    ('Cadena de frío: neveras, acumuladores y termómetro', 'Transporte', 'Logística', 'D · mañana'),
    ('Hoja de ruta con dirección, contacto y hora de llegada', 'Transporte', 'Logística', 'D-1'),

    ('Toma de corriente y potencia disponible verificadas', 'Montaje', 'Producción', 'D-3'),
    ('Punto de agua y evacuación localizados', 'Montaje', 'Producción', 'D-3'),
    ('Plano de sala: buffet, barras, office y circulación', 'Montaje', 'Jefe de sala', 'D-3'),

    ('Ficha de alérgenos de cada elaboración, impresa', 'Alérgenos', 'Jefe de cocina', 'D-1'),
    ('Comensales con alergia identificados y menú alternativo', 'Alérgenos', 'Jefe de cocina', 'D-1'),
    ('Etiquetado de alérgenos en las bandejas del buffet', 'Alérgenos', 'Jefe de sala', 'D · H-3 h'),

    ('Autorización del espacio y horario de música', 'Permisos', 'Comercial', 'D-7'),
    ('Seguro de responsabilidad civil en vigor', 'Permisos', 'Administración', 'D-7'),
    ('Registro sanitario y carnés de manipulador al día', 'Permisos', 'Administración', 'D-7'),
]

# ==========================================================================
# 08 · Punto de equilibrio (DOM-07/TEC-06/COM-03)
# ==========================================================================
# Concepto, importe €/día. Suman 300 €/día, que es el orden de magnitud de un
# food truck de plaza fija en España (canon de mercado + jornada de dos
# personas + generador). Todos son celdas verdes: el usuario pone los suyos.
COSTES_FIJOS = [
    ('Alquiler de plaza / canon del mercado', 60),
    ('Seguros, licencias y tasas (prorrateo diario)', 25),
    ('Combustible del truck y generador', 35),
    ('Personal (2 personas × jornada)', 140),
    ('Amortización del truck y del equipamiento', 30),
    ('Limpieza, consumibles y varios', 10),
]
MIX_08 = {'Smash Burger': 0.50, 'Loaded Fries': 0.25, 'Pulled Pork Sándwich': 0.25}


# ==========================================================================
# Utilidades
# ==========================================================================
def _verde(cel, fmt=None):
    cel.fill = PatternFill('solid', fgColor=motor.VERDE)
    cel.protection = Protection(locked=False)
    if fmt:
        cel.number_format = fmt
    return cel


def _hoja(wb, nombre, despues=None):
    """Hoja nueva (o la existente, vaciada por completo).

    Vaciar y reescribir entera es lo que hace idempotente a este módulo: la
    2.ª pasada no puede «acumular» nada porque no queda nada de la 1.ª.
    """
    nueva = nombre not in wb.sheetnames
    ws = wb.create_sheet(nombre) if nueva else wb[nombre]
    if nueva and despues and despues in wb.sheetnames:
        wb._sheets.remove(ws)
        wb._sheets.insert(wb._sheets.index(wb[despues]) + 1, ws)
    # PRIMERO deshacer las combinaciones: mientras una celda pertenece a un
    # rango combinado es un MergedCell y su `value` es de sólo lectura.
    for m in [str(x) for x in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for row in ws.iter_rows():
        for c in row:
            c.value = None
    ws.data_validations.dataValidation = []
    ws.conditional_formatting = ConditionalFormattingList()
    return ws, nueva


def _titulo(ws, fila, texto, hasta_col, col_ini=1, size=13):
    c = ws.cell(row=fila, column=col_ini, value=texto)
    c.font = Font(bold=True, size=size)
    c.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila,
                   end_column=hasta_col)
    return c


def _nota(ws, fila, texto, hasta_col, col_ini=1):
    c = ws.cell(row=fila, column=col_ini, value=texto)
    c.font = Font(size=9, italic=True)
    c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila,
                   end_column=hasta_col)
    return c


def _cabecera(ws, fila, col_ini, textos):
    for i, t in enumerate(textos):
        c = ws.cell(row=fila, column=col_ini + i, value=t)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=motor.CAB)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)


def _anchos(ws, col_ini, anchos):
    for i, a in enumerate(anchos):
        ws.column_dimensions[get_column_letter(col_ini + i)].width = a


def _subtitulo(ws, fila, texto, col_ini, col_fin):
    """Franja de sección. El relleno se pinta ANTES de combinar: después, las
    celdas de la derecha son MergedCell y ya no admiten asignación."""
    for col in range(col_ini, col_fin + 1):
        ws.cell(row=fila, column=col).fill = PatternFill('solid',
                                                         fgColor=motor.CREMA)
    c = ws.cell(row=fila, column=col_ini, value=texto)
    c.font = Font(bold=True, size=10)
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila,
                   end_column=col_fin)


def _oro(cel, fmt=None):
    cel.fill = PatternFill('solid', fgColor=motor.ORO)
    cel.font = Font(bold=True)
    if fmt:
        cel.number_format = fmt
    return cel


def _renombrar(ws, b, nuevos, informe):
    """Renombra rótulos del bloque de resultado. `nuevos` = {clave: etiqueta}.

    Las etiquetas nuevas tienen que estar en `motor.ALIAS_ROTULOS` o la 2.ª
    pasada del motor no las reconocerá.
    """
    for clave, etiqueta in nuevos.items():
        if clave not in b:
            informe.append(f'AVISO {ws.title}: no encuentro la fila «{clave}»')
            continue
        assert motor._canon(etiqueta) in (
            'COSTE TOTAL DEL PLATO', 'Coste elaboración (%)',
            motor.ETIQ_RACIONES, motor.ETIQ_COSTE_RACION,
            'PVP SUGERIDO (sin IVA)', 'PVP CON IVA'), etiqueta
        ws.cell(row=b[clave], column=1).value = etiqueta


def _fotos(ws, titulo, arrastre):
    """Los rótulos de la zona de foto («Foto del Plato») se buscan por CONTENIDO.

    El motor inserta la columna Factor en la 7, así que la columna de la foto
    se desplaza de K a L: buscarla por coordenada fija da en el sitio
    equivocado en la 2.ª pasada.
    """
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            if v.startswith('📷 Foto del') or v.startswith('📷 Foto de la'):
                c.value = titulo
            elif v.startswith('Arrastra aquí la foto'):
                c.value = arrastre


def _fila_por(ws, prefijo, col=2):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.startswith(prefijo):
            return r
    return None


def _reg(registro, ws, coord, valor):
    registro.append((ws.title, coord, valor))


# ==========================================================================
# API
# ==========================================================================
def pre(wb, fname, informe):
    """Único trabajo previo al motor: renombrar «Cocktail 50 pax» (06).

    Ver la cabecera del módulo: si se renombrase en post(), motor.REGISTRO ya
    habría anotado las fórmulas con el título viejo y la verificación de cache
    de main.py las daría por «hoja ausente».
    """
    if fname == '06-catering.xlsx' and HOJA_06_VIEJA in wb.sheetnames:
        wb[HOJA_06_VIEJA].title = HOJA_06
        informe.append(f'06: pestaña «{HOJA_06_VIEJA}» → «{HOJA_06}» '
                       '(TEC-07/COM-19: contenía cantidades POR PERSONA y el '
                       'Presupuesto volvía a multiplicar por los comensales)')


def post(wb, fname, informe, registro):
    if fname == '04-cocktails-bebidas.xlsx':
        _cocteles(wb, informe, registro)
    elif fname == '05-pasteleria.xlsx':
        _pasteleria(wb, informe, registro)
    elif fname == '06-catering.xlsx':
        _catering(wb, informe, registro)
    elif fname == '08-food-truck.xlsx':
        _food_truck(wb, informe, registro)


# --------------------------------------------------------------------------
# 04 · Cócteles
# --------------------------------------------------------------------------
def _cocteles(wb, informe, registro):
    for ws, _ in motor.hojas_escandallo(wb):
        b = motor.bloque(ws)
        _renombrar(ws, b, {'plato': 'COSTE TOTAL DEL CÓCTEL',
                           'elaboracion': 'Merma y hielo (%)'}, informe)
        cel = ws.cell(row=b['elaboracion'], column=9)
        cel.value = MERMA_HIELO
        cel.number_format = motor.FMT_PCT
        _verde(cel, motor.FMT_PCT)
        ws['A2'] = ('Cantidades en la unidad de USO (cl los destilados) y '
                    'precio en la de COMPRA (€/L). Las celdas verdes son las '
                    'editables.')
        _fotos(ws, '📷 Foto del Cóctel',
               'Arrastra aquí la foto\ndel cóctel terminado')
    informe.append('04: rótulos de cóctel («COSTE TOTAL DEL CÓCTEL», «Merma y '
                   f'hielo (%)» = {MERMA_HIELO:.0%}) en las 4 pestañas')
    _formatos_compra(wb, informe, registro)


def _formatos_compra(wb, informe, registro):
    ws, nueva = _hoja(wb, 'Formatos de Compra', despues='Instrucciones')
    _anchos(ws, 1, [30, 22, 24, 22, 30])
    _titulo(ws, 1, 'Del precio de la botella al precio por litro', 5)
    _nota(ws, 2,
          'La columna «Precio/Ud (€)» del escandallo pide €/LITRO, pero la '
          'factura del proveedor viene por botella. Escribe aquí el formato y '
          'lo que pagas por botella, y copia el «Precio por litro» a la '
          'columna «Precio/Ud» de la pestaña del cóctel. Los destilados se '
          'compran en botella de 70 cl y el vino y el espumoso en 75 cl: '
          'teclear el precio de la botella tal cual infravalora el coste un '
          '30 %.', 5)
    _cabecera(ws, 4, 1, ['Producto', 'Formato de compra (cl)',
                         'Precio de la botella (€)', 'Precio por litro (€)',
                         'Dónde se usa'])
    fila = 5
    for nombre, formato, precio, uso in BOTELLAS:
        ws.cell(row=fila, column=1, value=nombre)
        _verde(ws.cell(row=fila, column=2, value=formato), FMT_ENT).alignment = \
            Alignment(horizontal='center')
        _verde(ws.cell(row=fila, column=3, value=precio), motor.FMT_EUR)
        # TEC-R2-15: este valor ya no es informativo, alimenta el
        # «Precio/Ud» de los escandallos: sin ROUND, 19,60/0,70 imprimía
        # 28,000000000000004 en la columna de precios del cóctel.
        f = f'=IFERROR(ROUND(C{fila}*100/B{fila},4),"")'
        c = ws.cell(row=fila, column=4, value=f)
        c.number_format = motor.FMT_EUR
        c.font = Font(bold=True)
        _reg(registro, ws, f'D{fila}', f)
        ws.cell(row=fila, column=5, value=uso).font = Font(size=9, italic=True)
        fila += 1
    for _ in range(BOTELLAS_LIBRES):
        _verde(ws.cell(row=fila, column=1))
        _verde(ws.cell(row=fila, column=2), FMT_ENT).alignment = \
            Alignment(horizontal='center')
        _verde(ws.cell(row=fila, column=3), motor.FMT_EUR)
        f = f'=IFERROR(ROUND(C{fila}*100/B{fila},4),"")'
        c = ws.cell(row=fila, column=4, value=f)
        c.number_format = motor.FMT_EUR
        c.font = Font(bold=True)
        _reg(registro, ws, f'D{fila}', f)
        fila += 1
    _nota(ws, fila + 1,
          'El «Precio/Ud (€)» de las pestañas de cóctel está ENLAZADO a la '
          'columna «Precio por litro» de esta hoja (celdas azules): cambia '
          'aquí el precio de la botella y el escandallo se recalcula solo. Una '
          'ginebra de 19,60 € la botella de 70 cl son 28 €/L, que es lo que '
          'aparece en «Gin Tonic Premium».', 5)
    _enlazar_precios(wb, informe, registro)
    motor.print_setup(ws, 4, landscape=False)
    if nueva:
        informe.append('04: hoja «Formatos de Compra» creada '
                       f'({len(BOTELLAS)} referencias + {BOTELLAS_LIBRES} '
                       'libres; DOM-22)')


def _enlazar_precios(wb, informe, registro):
    """TEC-R2-15: «Precio/Ud» de destilados, vino y espumoso ← Formatos de
    Compra!D. Quedan BLOQUEADAS y en azul, no en verde: el verde significa «lo
    escribes tú» en todo el kit y aquí el dato viene de otra hoja."""
    hechos = []
    for idx, (hoja, fila_esc) in sorted(ENLACES_PRECIO.items()):
        if hoja not in wb.sheetnames:
            continue
        fila_fmt = 5 + idx
        cel = wb[hoja].cell(row=fila_esc, column=4)
        f = f"=IFERROR('Formatos de Compra'!$D${fila_fmt},\"\")"
        cel.value = f
        cel.number_format = motor.FMT_EUR
        cel.fill = PatternFill('solid', fgColor=AZUL_ENLACE)
        cel.protection = Protection(locked=True)
        _reg(registro, wb[hoja], f'D{fila_esc}', f)
        hechos.append(f'{hoja}!D{fila_esc}←Formatos!D{fila_fmt}')
    informe.append('04: «Precio/Ud» enlazado a «Formatos de Compra» en '
                   f'{len(hechos)} filas ({", ".join(hechos)}); celdas azules '
                   'y bloqueadas (TEC-R2-15)')


# --------------------------------------------------------------------------
# 05 · Pastelería
# --------------------------------------------------------------------------
def _pasteleria(wb, informe, registro):
    for ws, _ in motor.hojas_escandallo(wb):
        b = motor.bloque(ws)
        _renombrar(ws, b, {
            'plato': 'COSTE TOTAL DE LA TANDA',
            'raciones': 'Rendimiento (uds por tanda)',
            'coste_racion': 'COSTE POR UNIDAD',
            'pvp_sin': 'PVP POR UNIDAD (sin IVA)',
            'pvp_con': 'PVP POR UNIDAD (con IVA)'}, informe)
        rend = RENDIMIENTO_05.get(ws.title)
        if rend:
            cel = ws.cell(row=b['raciones'], column=9)
            cel.value = rend
            cel.number_format = FMT_ENT
            _verde(cel, FMT_ENT)
            _reg(registro, ws, f"I{b['raciones']}", rend)
        for fila, merma in MERMAS_05.get(ws.title, {}).items():
            ws.cell(row=fila, column=8).value = merma
            ws.cell(row=fila, column=8).number_format = motor.FMT_PCT
        ws['A2'] = ('El escandallo es de la TANDA COMPLETA: el «Rendimiento» '
                    'reparte el coste entre las unidades que salen. Las '
                    'celdas verdes son las editables.')
        _fotos(ws, '📷 Foto de la Elaboración',
               'Arrastra aquí la foto\nde la elaboración terminada')
    informe.append('05: rendimiento por tanda '
                   + ', '.join(f'{h} {n}' for h, n in RENDIMIENTO_05.items())
                   + '; PVP y coste POR UNIDAD (DOM-04/TEC-05/COM-01)')
    informe.append('05: mermas de obrador escritas — '
                   'Tarta Chocolate!H5 12 % (temperado), Tarta Chocolate!H9 y '
                   'Croissants!H5 8 % (horneado), Macarons!H10 12 % '
                   '(TEC-20/COM-13)')


# --------------------------------------------------------------------------
# 06 · Catering
# --------------------------------------------------------------------------
def _catering(wb, informe, registro):
    ws = wb[HOJA_06]
    b = motor.bloque(ws)
    _renombrar(ws, b, {'plato': 'COSTE TOTAL POR PERSONA'}, informe)
    ws['A1'] = 'Catering Cocktail — Coste por Persona'
    ws['A2'] = ('AVISO: todas las cantidades son POR COMENSAL. El número de '
                'comensales se escribe UNA sola vez, en «Presupuesto»!C5 — no '
                'lo multipliques aquí o facturarás el evento dos veces.')
    _fotos(ws, '📷 Foto del Montaje',
           'Arrastra aquí la foto\ndel montaje terminado')
    # TEC-08: el bloque de PVP de esta hoja compite con el del Presupuesto.
    nota = ws.cell(row=b['fc_real'] + 2, column=1)
    nota.value = ('PVP orientativo: cubre sólo la materia prima. El precio que '
                  'se le da al cliente sale de la hoja «Presupuesto», que '
                  'suma personal, menaje, transporte y montaje.')
    nota.font = Font(size=9, italic=True)
    filas = _presupuesto_06(wb, b, informe, registro)
    _checklist_06(wb, informe, registro)
    _presupuesto_cliente_06(wb, informe, registro, filas)


def _presupuesto_06(wb, b_esc, informe, registro):
    """Reconstruye «Presupuesto» entero (DOM-12/COM-05/DOM-19/TEC-21).

    Los estilos se capturan por ETIQUETA, nunca por número de fila: la hoja
    pasa de 20 filas a 29 y en la 2.ª pasada la fila 13 ya no es el total.
    """
    ws = wb['Presupuesto']
    f_hdr = _fila_por(ws, 'Concepto') or 4
    f_lbl = _fila_por(ws, 'Número de comensales') or 5
    f_tot = _fila_por(ws, 'COSTE TOTAL') or 13
    f_in = _fila_por(ws, 'Food Cost objetivo') or _fila_por(ws, 'Margen objetivo') or 16
    f_pvp = _fila_por(ws, 'PVP POR PERSONA') or 17
    est = {
        'hdr': (copy.copy(ws.cell(row=f_hdr, column=2)._style),
                copy.copy(ws.cell(row=f_hdr, column=3)._style)),
        'lbl': (copy.copy(ws.cell(row=f_lbl, column=2)._style),
                copy.copy(ws.cell(row=f_lbl, column=3)._style)),
        'tot': (copy.copy(ws.cell(row=f_tot, column=2)._style),
                copy.copy(ws.cell(row=f_tot, column=3)._style)),
        'in': (copy.copy(ws.cell(row=f_in, column=2)._style),
               copy.copy(ws.cell(row=f_in, column=3)._style)),
        'pvp': (copy.copy(ws.cell(row=f_pvp, column=2)._style),
                copy.copy(ws.cell(row=f_pvp, column=3)._style)),
    }
    est_tit = copy.copy(ws.cell(row=2, column=2)._style)

    # valores del cliente que se conservan entre pasadas
    previos = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip():
            previos[v.strip()] = ws.cell(row=r, column=3).value

    for m in [str(x) for x in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for row in ws.iter_rows():
        for c in row:
            c.value = None
    ws.data_validations.dataValidation = []
    ws.conditional_formatting = ConditionalFormattingList()

    def _num(etiqueta, defecto):
        v = previos.get(etiqueta)
        return v if isinstance(v, (int, float)) else defecto

    def _fila(fila, etiqueta, banda, valor=None, formula=None, fmt=None,
              verde=False):
        cl = ws.cell(row=fila, column=2, value=etiqueta)
        cl._style = copy.copy(est[banda][0])
        cv = ws.cell(row=fila, column=3)
        cv._style = copy.copy(est[banda][1])
        if formula is not None:
            cv.value = formula
            _reg(registro, ws, f'C{fila}', formula)
        elif valor is not None:
            cv.value = valor
        if fmt:
            cv.number_format = fmt
        if verde:
            _verde(cv, fmt)
        elif formula is not None:
            # DOM-10: nunca una fórmula en verde.
            cv.fill = PatternFill(fill_type=None)
            cv.protection = Protection(locked=True)
        return cv

    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 18

    c = ws.cell(row=2, column=2, value='Presupuesto de Catering — hoja INTERNA')
    c._style = copy.copy(est_tit)
    n3 = ws.cell(row=3, column=2,
                 value='Enseña tus costes y tu margen: NO se le manda al '
                       'cliente. Para eso está «Presupuesto Cliente».')
    n3.font = Font(size=9, italic=True)

    _fila(4, 'Concepto', 'hdr')
    ws.cell(row=4, column=3, value='Importe (€)')._style = copy.copy(est['hdr'][1])

    _fila(5, 'Número de comensales', 'lbl',
          valor=int(_num('Número de comensales', 50)), fmt=FMT_ENT, verde=True)
    # TEC-R2-06: leía «COSTE TOTAL POR PERSONA» (la fila del total de la
    # tanda). Con «Nº de raciones» > 1 —que es justo lo que la fila nueva de
    # §1.4 invita a escribir— el presupuesto facturaba la tanda entera por
    # comensal. La celda correcta es COSTE POR RACIÓN, la misma de la que
    # cuelga el PVP de la propia pestaña.
    # TEC-R2-05/COM-M14: es una FÓRMULA — va bloqueada, sin verde y rotulada
    # como enlace, no como campo a rellenar.
    _fila(6, 'Coste de alimentación por comensal (€) — enlazado al escandallo '
             f'«{HOJA_06}»', 'lbl',
          formula=f"='{HOJA_06}'!$J${b_esc['coste_racion']}", fmt=motor.FMT_EUR)
    _fila(7, 'Coste total de alimentación (€)', 'lbl',
          formula='=IFERROR(C5*C6,"")', fmt=motor.FMT_EUR)

    _subtitulo(ws, 8, 'Personal de servicio', 2, 3)
    _fila(9, 'Comensales por camarero', 'lbl',
          valor=_num('Comensales por camarero', 22), fmt=FMT_ENT, verde=True)
    _fila(10, 'Nº de camareros (calculado)', 'lbl',
          formula='=IFERROR(ROUNDUP(C5/C9,0),"")', fmt=FMT_ENT)
    _fila(11, 'Horas por camarero', 'lbl',
          valor=_num('Horas por camarero', 5), fmt=FMT_H, verde=True)
    _fila(12, 'Coste por hora de camarero (€)', 'lbl',
          valor=_num('Coste por hora de camarero (€)', 16), fmt=motor.FMT_EUR,
          verde=True)
    _fila(13, 'Horas de jefe de sala', 'lbl',
          valor=_num('Horas de jefe de sala', 6), fmt=FMT_H, verde=True)
    _fila(14, 'Coste por hora de jefe de sala (€)', 'lbl',
          valor=_num('Coste por hora de jefe de sala (€)', 22),
          fmt=motor.FMT_EUR, verde=True)
    _fila(15, 'COSTE DE PERSONAL (€)', 'tot',
          formula='=IFERROR(C10*C11*C12+C13*C14,"")', fmt=motor.FMT_EUR)

    _subtitulo(ws, 16, 'Menaje, transporte y montaje', 2, 3)
    _fila(17, 'Menaje y desechables por comensal (€)', 'lbl',
          valor=_num('Menaje y desechables por comensal (€)', 1.20),
          fmt=motor.FMT_EUR, verde=True)
    _fila(18, 'Coste de menaje (€)', 'lbl',
          formula='=IFERROR(C5*C17,"")', fmt=motor.FMT_EUR)
    _fila(19, 'Transporte y logística (€)', 'lbl',
          valor=_num('Transporte y logística (€)', 150), fmt=motor.FMT_EUR,
          verde=True)
    _fila(20, 'Montaje y desmontaje (€)', 'lbl',
          valor=_num('Montaje y desmontaje (€)', 200), fmt=motor.FMT_EUR,
          verde=True)

    _fila(21, 'COSTE TOTAL DEL EVENTO (€)', 'tot',
          formula='=IFERROR(C7+C15+C18+C19+C20,"")', fmt=motor.FMT_EUR)
    _fila(22, 'Coste por comensal (€)', 'tot',
          formula='=IFERROR(C21/C5,"")', fmt=motor.FMT_EUR)

    _subtitulo(ws, 23, 'Precio de venta', 2, 3)
    _fila(24, 'Food Cost objetivo de la COMIDA (%)', 'in',
          valor=_num('Food Cost objetivo de la COMIDA (%)',
                     _num('Food Cost objetivo (%)', 0.35)),
          fmt=motor.FMT_PCT, verde=True)
    _fila(25, 'Margen sobre servicios (%)', 'in',
          valor=_num('Margen sobre servicios (%)', 0.20), fmt=motor.FMT_PCT,
          verde=True)
    _fila(26, 'PVP de la alimentación (€)', 'lbl',
          formula='=IFERROR(C7/C24,"")', fmt=motor.FMT_EUR)
    _fila(27, 'PVP de los servicios (personal, menaje, transporte, montaje) (€)',
          'lbl', formula='=IFERROR((C15+C18+C19+C20)*(1+C25),"")',
          fmt=motor.FMT_EUR)
    _fila(28, 'Mínimo de facturación por evento (€)', 'in',
          valor=_num('Mínimo de facturación por evento (€)', 600),
          fmt=motor.FMT_EUR, verde=True)
    _fila(29, 'PVP DEL EVENTO (sin IVA)', 'pvp',
          formula='=IFERROR(MAX(C26+C27,C28),"")', fmt=motor.FMT_EUR)
    _fila(30, 'PVP POR PERSONA (sin IVA)', 'pvp',
          formula='=IFERROR(C29/C5,"")', fmt=motor.FMT_EUR)
    _fila(31, 'Tipo de IVA (%)', 'in', valor=_num('Tipo de IVA (%)', 0.10),
          fmt=motor.FMT_PCT, verde=True)
    _fila(32, 'PVP POR PERSONA (con IVA)', 'pvp',
          formula='=IFERROR(C30*(1+C31),"")', fmt=motor.FMT_EUR)
    _fila(33, 'PRESUPUESTO TOTAL (con IVA)', 'tot',
          formula='=IFERROR(C29*(1+C31),"")', fmt=motor.FMT_EUR)
    _fila(34, 'Margen bruto del evento (€)', 'lbl',
          formula='=IFERROR(C29-C21,"")', fmt=motor.FMT_EUR)
    _fila(35, 'Food cost de la comida sobre lo facturado (%)', 'lbl',
          formula='=IFERROR(C7/C29,"")', fmt=motor.FMT_PCT)

    n = ws.cell(row=37, column=2,
                value='El food cost objetivo se aplica SÓLO a la comida '
                      '(C7 ÷ C24). El personal, el menaje, el transporte y el '
                      'montaje no llevan food cost: van a coste más el margen '
                      'de servicios de C25. Aplicar el 35 % al coste total '
                      'multiplicaba el presupuesto por 2,9 y perdías el '
                      'evento. C35 te enseña qué food cost de comida te queda '
                      'sobre el total facturado. Ojo con los eventos '
                      'pequeños: el transporte, el montaje y el jefe de sala '
                      'son fijos, así que con 10 comensales el precio por '
                      'persona se dispara por encima de los 100 €. No es un '
                      'error de la hoja: es lo que cuesta de verdad mover un '
                      'catering para diez. Ajusta esas tres celdas verdes a '
                      'la realidad de tu evento pequeño.')
    n.font = Font(size=9, italic=True)
    n.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=37, start_column=2, end_row=37, end_column=3)
    ws.row_dimensions[37].height = 116
    motor.print_setup(ws, 4, landscape=False)
    informe.append('06: «Presupuesto» reconstruido — food cost SÓLO sobre la '
                   'alimentación + margen de servicios + mínimo de '
                   'facturación (SPEC §7.2 / DOM-03), personal por camareros × '
                   'horas × €/h con ROUNDUP(pax/22) y menaje por comensal en '
                   'celda verde (DOM-12/COM-05/DOM-19/TEC-21)')
    return {'pvp_persona_iva': 32, 'pvp_evento_iva': 33, 'pax': 5}


def _checklist_06(wb, informe, registro):
    ws, nueva = _hoja(wb, 'Checklist Evento', despues='Presupuesto')
    _anchos(ws, 1, CHK_ANCHOS)
    ancho = len(CHK_HDR)
    _titulo(ws, 1, 'Checklist de Evento — Catering', ancho)
    _nota(ws, 2,
          'Marca con ✓ en la columna «✓ Completada» (es un desplegable: '
          '✓ / — / N/A). El contador del final cuenta sólo los ✓. Las cuatro '
          'filas en blanco del final son para tus propias tareas y también '
          'cuentan.', ancho)
    _cabecera(ws, 4, 1, CHK_HDR)

    fila = 5
    for i, (tarea, bloque, resp, cuando) in enumerate(CHECKLIST, start=1):
        ws.cell(row=fila, column=1, value=i).alignment = \
            Alignment(horizontal='center')
        ws.cell(row=fila, column=2, value=tarea).alignment = \
            Alignment(wrap_text=True, vertical='center')
        ws.cell(row=fila, column=3, value=bloque)
        ws.cell(row=fila, column=4, value=resp)
        ws.cell(row=fila, column=5, value=cuando).alignment = \
            Alignment(horizontal='center')
        _verde(ws.cell(row=fila, column=6)).alignment = \
            Alignment(horizontal='center')
        fila += 1
    ultima_tarea = fila - 1
    for _ in range(HOLGURA):
        _verde(ws.cell(row=fila, column=2))
        _verde(ws.cell(row=fila, column=6)).alignment = \
            Alignment(horizontal='center')
        fila += 1
    fin = fila - 1

    dv = DataValidation(type='list', formula1=DV_MARCA, allow_blank=True,
                        showDropDown=False)
    dv.showErrorMessage = True
    dv.errorTitle = 'Valor no válido'
    dv.error = 'Elige ✓, — o N/A.'
    ws.add_data_validation(dv)
    dv.add(f'F5:F{fin}')

    ws.conditional_formatting.add(
        f'A5:F{fin}',
        FormulaRule(formula=['$F5="✓"'],
                    fill=PatternFill('solid', start_color='C6EFCE',
                                     end_color='C6EFCE')))

    r_cnt = fin + 2
    c = ws.cell(row=r_cnt, column=1, value='Tareas completadas:')
    c.font = Font(bold=True)
    ws.merge_cells(start_row=r_cnt, start_column=1, end_row=r_cnt, end_column=3)
    f_num = f'=COUNTIF(F5:F{fin},"✓")'
    cn = _oro(ws.cell(row=r_cnt, column=4, value=f_num), FMT_ENT)
    cn.alignment = Alignment(horizontal='center')
    _reg(registro, ws, f'D{r_cnt}', f_num)
    ws.cell(row=r_cnt, column=5, value='de').alignment = \
        Alignment(horizontal='center')
    f_den = f'=COUNTIF(B5:B{fin},"?*")'
    cd = _oro(ws.cell(row=r_cnt, column=6, value=f_den), FMT_ENT)
    cd.alignment = Alignment(horizontal='center')
    _reg(registro, ws, f'F{r_cnt}', f_den)

    # TEC-R2-12: los dos campos de firma no tenían NINGUNA celda desbloqueada
    # al lado en una hoja protegida — el único hueco del libro que el diseño
    # pide rellenar y la protección impedía escribir. Se abren dos columnas
    # (B y C) por fila y la fecha lleva su formato.
    ws.cell(row=r_cnt + 2, column=1, value='Responsable del evento:').font = \
        Font(bold=True)
    for col in (2, 3):
        _verde(ws.cell(row=r_cnt + 2, column=col))
    ws.cell(row=r_cnt + 3, column=1, value='Fecha del evento:').font = \
        Font(bold=True)
    for col in (2, 3):
        _verde(ws.cell(row=r_cnt + 3, column=col), 'dd/mm/yyyy')

    motor.print_setup(ws, 4, landscape=True)
    if nueva:
        informe.append(f'06: hoja «Checklist Evento» creada '
                       f'({len(CHECKLIST)} tareas + {HOLGURA} libres, '
                       'desplegable ✓/—/N/A y contador; TEC-24/COM-16)')


def _presupuesto_cliente_06(wb, informe, registro, filas=None):
    # DOM-03: la fila del «PVP POR PERSONA (con IVA)» se movió al separar el
    # food cost de la comida del margen de servicios. Se recibe, no se fija.
    filas = filas or {'pvp_persona_iva': 32, 'pax': 5}
    ws, nueva = _hoja(wb, 'Presupuesto Cliente', despues='Checklist Evento')
    _anchos(ws, 1, [4, 46, 14, 22, 18])
    _titulo(ws, 2, 'Propuesta de Catering', 5)

    for i, etiqueta in enumerate(('Cliente', 'Evento y fecha', 'Lugar'),
                                 start=3):
        ws.cell(row=i, column=2, value=etiqueta + ':').font = Font(bold=True)
        _verde(ws.cell(row=i, column=3))
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)

    _cabecera(ws, 7, 2, ['Concepto', 'Comensales',
                         'Precio por persona (IVA incl.)', 'Total (€)'])
    f = f"=IFERROR('Presupuesto'!C{filas['pax']},\"\")"
    # TEC-R2-04: el concepto de la primera línea es un DATO precargado, no un
    # rótulo de estructura: es lo primero que cada caterer reescribe con el
    # nombre de su menú. Iba bloqueado y en gris mientras B9:B11 sí eran
    # verdes — dos criterios en la misma columna, y en la única hoja que se le
    # enseña al cliente.
    c = ws.cell(row=8, column=2,
                value='Servicio de catering cocktail — menú acordado')
    _verde(c)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    cc = ws.cell(row=8, column=3, value=f)
    cc.number_format = FMT_ENT
    cc.alignment = Alignment(horizontal='center')
    _reg(registro, ws, 'C8', f)
    f = f"=IFERROR('Presupuesto'!C{filas['pvp_persona_iva']},\"\")"
    cd = ws.cell(row=8, column=4, value=f)
    cd.number_format = motor.FMT_EUR
    _reg(registro, ws, 'D8', f)
    f = '=IF(B8="","",IFERROR(C8*D8,0))'
    ce = ws.cell(row=8, column=5, value=f)
    ce.number_format = motor.FMT_EUR
    _reg(registro, ws, 'E8', f)

    for r in (9, 10, 11):
        _verde(ws.cell(row=r, column=2))
        _verde(ws.cell(row=r, column=3), FMT_ENT).alignment = \
            Alignment(horizontal='center')
        _verde(ws.cell(row=r, column=4), motor.FMT_EUR)
        f = f'=IF(B{r}="","",IFERROR(C{r}*D{r},0))'
        c = ws.cell(row=r, column=5, value=f)
        c.number_format = motor.FMT_EUR
        _reg(registro, ws, f'E{r}', f)

    ws.cell(row=13, column=2, value='TOTAL (IVA incluido)').font = \
        Font(bold=True, size=12)
    f = '=IFERROR(SUM(E8:E11),"")'
    ct = _oro(ws.cell(row=13, column=5, value=f), motor.FMT_EUR)
    _reg(registro, ws, 'E13', f)

    _nota(ws, 15,
          'Presupuesto válido 30 días desde su emisión. Precios con IVA '
          'incluido. El número definitivo de comensales se confirma 7 días '
          'antes del evento; a partir de esa fecha se factura el número '
          'confirmado.', 5)
    ws.cell(row=17, column=2, value=motor.PIE).font = Font(size=9)
    motor.print_setup(ws, None, landscape=False)
    if nueva:
        informe.append('06: hoja «Presupuesto Cliente» creada (concepto, '
                       'comensales y PVP con IVA; sin costes ni margen a la '
                       'vista; TEC-24/COM-16)')


# --------------------------------------------------------------------------
# 08 · Food truck
# --------------------------------------------------------------------------
def _food_truck(wb, informe, registro):
    # Los importes que haya escrito el cliente se leen ANTES de vaciar la hoja
    # (si se leyeran después, _hoja ya los habría borrado).
    previos = {}
    if 'Punto de Equilibrio' in wb.sheetnames:
        w0 = wb['Punto de Equilibrio']
        for r in range(1, w0.max_row + 1):
            v = w0.cell(row=r, column=1).value
            if isinstance(v, str):
                previos[v] = w0.cell(row=r, column=2).value
    ws, nueva = _hoja(wb, 'Punto de Equilibrio',
                      despues='Pulled Pork Sándwich')
    _anchos(ws, 1, [42, 18, 18, 18, 16])
    _titulo(ws, 1, 'Punto de Equilibrio Diario — Food Truck', 5)
    _nota(ws, 2,
          'Cuántas unidades tienes que vender cada día sólo para cubrir los '
          'costes fijos. El margen sale de las pestañas de escandallo: si '
          'cambias una receta o un precio, esta hoja se mueve sola. El PVP es '
          'el que tengas escrito en «PVP actual en carta (sin IVA)» de cada '
          'pestaña; mientras esa celda esté vacía se usa el PVP sugerido, que '
          'casi nunca es el que cobras.', 5)

    _subtitulo(ws, 4, 'Costes fijos del día', 1, 5)
    _cabecera(ws, 5, 1, ['Concepto', 'Importe (€/día)'])
    fila = 6
    for concepto, importe in COSTES_FIJOS:
        ws.cell(row=fila, column=1, value=concepto)
        v = previos.get(concepto)
        _verde(ws.cell(row=fila, column=2,
                       value=v if isinstance(v, (int, float)) else importe),
               motor.FMT_EUR)
        fila += 1
    r_fijos = fila
    ws.cell(row=r_fijos, column=1, value='COSTES FIJOS TOTALES (€/día)').font = \
        Font(bold=True)
    f = f'=SUM(B6:B{r_fijos - 1})'
    _oro(ws.cell(row=r_fijos, column=2, value=f), motor.FMT_EUR)
    _reg(registro, ws, f'B{r_fijos}', f)

    r_mix = r_fijos + 2
    _subtitulo(ws, r_mix, 'Margen de contribución por unidad', 1, 5)
    _cabecera(ws, r_mix + 1, 1,
              ['Producto', 'PVP sin IVA (€)', 'Coste unitario (€)',
               'Margen (€)', 'Mix de ventas'])
    fila = r_mix + 2
    p0 = fila
    for titulo, mix in MIX_08.items():
        b = motor.bloque(wb[titulo])
        ws.cell(row=fila, column=1, value=titulo)
        f = (f"=IF('{titulo}'!$I${b['pvp_actual']}=\"\","
             f"'{titulo}'!$J${b['pvp_sin']},'{titulo}'!$I${b['pvp_actual']})")
        c = ws.cell(row=fila, column=2, value=f)
        c.number_format = motor.FMT_EUR
        _reg(registro, ws, f'B{fila}', f)
        f = f"='{titulo}'!$J${b['coste_racion']}"
        c = ws.cell(row=fila, column=3, value=f)
        c.number_format = motor.FMT_EUR
        _reg(registro, ws, f'C{fila}', f)
        f = f'=IFERROR(B{fila}-C{fila},"")'
        c = ws.cell(row=fila, column=4, value=f)
        c.number_format = motor.FMT_EUR
        c.font = Font(bold=True)
        _reg(registro, ws, f'D{fila}', f)
        _verde(ws.cell(row=fila, column=5, value=mix), motor.FMT_PCT).alignment = \
            Alignment(horizontal='center')
        fila += 1
    p1 = fila - 1

    r_pvpm = fila
    ws.cell(row=r_pvpm, column=1, value='PVP MEDIO PONDERADO (€/ud)').font = \
        Font(bold=True)
    f = (f'=IFERROR(SUMPRODUCT(B{p0}:B{p1},E{p0}:E{p1})/SUM(E{p0}:E{p1}),"")')
    c = ws.cell(row=r_pvpm, column=2, value=f)
    c.number_format = motor.FMT_EUR
    c.font = Font(bold=True)
    _reg(registro, ws, f'B{r_pvpm}', f)

    r_marg = fila + 1
    ws.cell(row=r_marg, column=1,
            value='MARGEN MEDIO PONDERADO (€/ud)').font = Font(bold=True)
    f = (f'=IFERROR(SUMPRODUCT(D{p0}:D{p1},E{p0}:E{p1})/SUM(E{p0}:E{p1}),"")')
    _oro(ws.cell(row=r_marg, column=2, value=f), motor.FMT_EUR)
    _reg(registro, ws, f'B{r_marg}', f)

    r_res = r_marg + 2
    _subtitulo(ws, r_res, 'Resultado', 1, 5)
    ws.cell(row=r_res + 1, column=1,
            value='UNIDADES/DÍA PARA CUBRIR COSTES').font = Font(bold=True)
    f = f'=IFERROR(ROUNDUP(B{r_fijos}/B{r_marg},0),"")'
    _oro(ws.cell(row=r_res + 1, column=2, value=f), FMT_ENT).alignment = \
        Alignment(horizontal='center')
    _reg(registro, ws, f'B{r_res + 1}', f)

    ws.cell(row=r_res + 2, column=1,
            value='FACTURACIÓN MÍNIMA sin IVA (€/día)').font = Font(bold=True)
    f = f'=IFERROR(B{r_res + 1}*B{r_pvpm},"")'
    c = ws.cell(row=r_res + 2, column=2, value=f)
    c.number_format = motor.FMT_EUR
    _reg(registro, ws, f'B{r_res + 2}', f)

    ws.cell(row=r_res + 3, column=1, value='Tipo de IVA (%)')
    prev_iva = previos.get('Tipo de IVA (%)')
    _verde(ws.cell(row=r_res + 3, column=2,
                   value=prev_iva if isinstance(prev_iva, (int, float)) else 0.10),
           motor.FMT_PCT).alignment = Alignment(horizontal='center')

    ws.cell(row=r_res + 4, column=1,
            value='FACTURACIÓN MÍNIMA con IVA (€/día)').font = Font(bold=True)
    f = f'=IFERROR(B{r_res + 2}*(1+B{r_res + 3}),"")'
    c = ws.cell(row=r_res + 4, column=2, value=f)
    c.number_format = motor.FMT_EUR
    _reg(registro, ws, f'B{r_res + 4}', f)

    _nota(ws, r_res + 6,
          'El «mix de ventas» es qué proporción de lo que vendes es cada '
          'producto: si un día vendes mitad de hamburguesas y el resto a '
          'partes iguales, deja 50 % / 25 % / 25 %. No tiene que sumar 100 %: '
          'la hoja pondera con lo que haya.', 5)
    motor.print_setup(ws, None, landscape=False)
    if nueva:
        informe.append('08: hoja «Punto de Equilibrio» creada (costes fijos '
                       'del día, margen medio ponderado de las 3 recetas y '
                       'unidades/día; DOM-07/TEC-06/COM-03)')
