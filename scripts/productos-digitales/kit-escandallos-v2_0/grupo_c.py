#!/usr/bin/env python3
"""
grupo_c.py — §4 de `kit-escandallos-v2-SPEC.md`: 09, 10, 11 y el BONUS.

  · 09 control de mermas (DOM-02/TEC-04/COM-22/DOM-09/TEC-03/COM-08/DOM-10/
    TEC-22/COM-20/COM-09/DOM-26): la columna de referencia deja de ser el
    RENDIMIENTO DE DESPIECE y pasa a ser «Desperdicio objetivo (% s/compra)»
    con mín / típico / máx por familia; semáforo REAL con formato condicional;
    hoja «Evolución» de 12 semanas alimentada por el TOTAL, con LineChart de
    openpyxl.
  · 10 calculadora PVP (DOM-11/TEC-18/COM-11): fila «Delivery» con comisión de
    plataforma descontada antes del PVP, columna «Multiplicador» = 1/media del
    rango, IVA en celda.
  · 11 dashboard (DOM-03/TEC-14/COM-10/TEC-13/COM-24/TEC-27/COM-27): stock
    inicial y final → food cost = consumo/ventas; rótulos «sin IVA»; columna
    «Estado» + formato condicional; serie objetivo condicionada; gráfico
    REGENERADO con openpyxl (categorías de TEXTO y eje abajo); rótulo del input
    sin cortar.
  · BONUS inventario (DOM-15/TEC-19/COM-31/COM-32/DOM-25/TEC-28): pestaña
    «Ventas del periodo» + consumo teórico con SUMPRODUCT, precio unitario y
    valor de la diferencia, guarda de vacío y registro de 90 filas.

PIPELINE: estos 4 ficheros NO pasan por `motor.aplicar` — no tienen rejilla de
escandallo, y `motor.reescribir_instrucciones()` les escribiría las
instrucciones de una hoja de escandallo (columnas Factor/Merma/Cant. Bruta que
aquí no existen) y `escribir_conversiones/mermas` les colgaría dos hojas
auxiliares ajenas. Por eso se declaran en `PROPIOS`: `main.py` salta el motor y
`post()` hace el trabajo entero, cerrando él mismo con `motor.cerrar()` (que sí
es genérico: limpieza de '', A4, protección y línea de versión).

IDEMPOTENTE por construcción: cada hoja de datos se RECONSTRUYE desde cero
(`_reset()` borra valores, estilos, merges, validaciones, formato condicional y
gráficos) y se vuelve a escribir con valores absolutos. No hay ningún paso que
dependa del estado anterior del fichero.
"""
import copy
import os
import shutil

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.datavalidation import DataValidation

import motor

F09 = '09-control-mermas.xlsx'
F10 = '10-calculadora-pvp.xlsx'
F11 = '11-dashboard-food-cost-mensual.xlsx'
FBO = 'BONUS-mermas-inventario.xlsx'

FICHEROS = [F09, F10, F11, FBO]
# main.py: no aplicar motor.aplicar/motor.cerrar; los gestiona este módulo.
PROPIOS = list(FICHEROS)

# ---- paleta ----------------------------------------------------------------
OK_BG, OK_FG = 'C6EFCE', '006100'
AMBAR = 'FFEB9C'
BANDA_A, BANDA_B = 'FFFFFF', 'F5F5F5'

FMT_EUR = motor.FMT_EUR                 # '#,##0.00 €'
FMT_EUR0 = '#,##0 €'
FMT_PCT1 = motor.FMT_PCT1               # '0.0%'
FMT_PP = '+0.0%;-0.0%'
FMT_UD = '0.00'
FMT_DUD = '+0.00;-0.00'
FMT_MULT = '0.00'
FMT_FECHA = 'dd/mm/yyyy'


# ==========================================================================
# §4.09 — Desperdicio objetivo por familia (mín / típico / máx, s/ COMPRA)
# ==========================================================================
# DOM-02/TEC-04: lo que había en la columna B eran los rendimientos de DESPIECE
# (carne 20 %, marisco 45 %), que ya están cobrados dentro del escandallo. Aquí
# se mide DESPERDICIO sobre la compra (caducidad, mal estado, sobreproducción),
# cuyo benchmark es un orden de magnitud menor. Familias de la SPEC §4:
# secos/congelados 2-3 %, carne/pescado 3-5 %, verdura/fruta 5-8 %, bebidas
# 1-2 %; el resto se sitúa por analogía y queda escrito en la propia hoja.
DESPERDICIO = [
    ('Carne roja',       0.030, 0.040, 0.050),
    ('Aves',             0.030, 0.040, 0.050),
    ('Pescado',          0.030, 0.040, 0.050),
    ('Marisco',          0.030, 0.050, 0.060),
    ('Verdura hoja',     0.050, 0.070, 0.080),
    ('Verdura raíz',     0.050, 0.060, 0.080),
    ('Fruta',            0.050, 0.060, 0.080),
    ('Lácteos',          0.020, 0.030, 0.040),
    ('Secos/granos',     0.020, 0.025, 0.030),
    ('Congelados',       0.020, 0.025, 0.030),
    ('Pan/bollería',     0.050, 0.080, 0.100),
    ('Huevos',           0.010, 0.020, 0.030),
    ('Aceites/grasas',   0.005, 0.010, 0.015),
    ('Especias/hierbas', 0.030, 0.050, 0.080),
    ('Chocolate/cacao',  0.010, 0.020, 0.030),
    ('Bebidas/licores',  0.010, 0.015, 0.020),
]

# Ejemplo cargado: dos categorías en ALERTA y tres en OK, para que el semáforo
# se vea al abrir (es lo que promete la landing y la propia hoja de
# instrucciones desde v1.1 sin cumplirlo — COM-08/TEC-03).
EJEMPLO_09 = {
    'Carne roja': (1850, 74),        # 4,0 % = objetivo → OK
    'Pescado': (1240, 87),           # 7,0 % > 4,0 %    → ALERTA
    'Verdura hoja': (620, 38),       # 6,1 % < 7,0 %    → OK
    'Lácteos': (480, 12),            # 2,5 % < 3,0 %    → OK
    'Pan/bollería': (310, 34),       # 11,0 % > 8,0 %   → ALERTA
}

# Semanas 2-4 del histórico (la 1 la trae el TOTAL de «Mermas Semanal»).
EJEMPLO_EVOL = [(4300, 205), (4150, 168), (4400, 154)]
OBJETIVO_EVOL = 0.04

# ==========================================================================
# §4.10 — Tipos de establecimiento
# ==========================================================================
# (etiqueta, food cost mín., food cost máx., comisión de plataforma)
# «Bar/Cócteles» pasa de 18-25 % a 20-25 % para que el kit tenga UN SOLO rango
# de food cost de bar (SPEC §3, 04-cocktails).
TIPOS_10 = [
    ('Fine Dining',      0.25, 0.28, 0.00),
    ('Casual Dining',    0.28, 0.32, 0.00),
    ('Fast Casual',      0.30, 0.35, 0.00),
    ('Cafetería',        0.25, 0.30, 0.00),
    ('Catering',         0.30, 0.40, 0.00),
    ('Food Truck',       0.28, 0.35, 0.00),
    ('Hotel F&B',        0.30, 0.35, 0.00),
    ('Pastelería',       0.20, 0.30, 0.00),
    ('Bar/Cócteles',     0.20, 0.25, 0.00),
    ('Delivery',         0.28, 0.32, 0.30),
]
COSTE_10 = 5.50
IVA_10 = 0.10

# ==========================================================================
# §4.11 — Dashboard
# ==========================================================================
MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
         'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
# (stock inicial, compras, stock final, ventas) — todo SIN IVA.
# Enero es exactamente el caso de la prueba: 1.000 + 5.000 − 1.200 = 4.800
# sobre 16.000 → 30,0 % → OK con el objetivo en 30 %.
EJEMPLO_11 = {
    'Enero': (1000, 5000, 1200, 16000),      # 30,0 % → OK
    'Febrero': (1200, 5400, 1100, 15000),    # 36,7 % → ALERTA
    'Marzo': (1100, 5200, 1300, 17000),      # 29,4 % → OK
}
OBJETIVO_11 = 0.30

# ==========================================================================
# §4.BONUS — Inventario
# ==========================================================================
# (producto, unidad, stock inicial, compras, stock final, precio unitario)
# El stock sólo va cargado en los 8 productos que aparecen en «Ventas del
# periodo»: si se cargara en los 15, los 7 sin ventas mostrarían todo su
# consumo como «pérdida oculta», que es exactamente lo contrario de lo que
# tiene que enseñar la hoja.
INVENTARIO = [
    ('Solomillo ternera', 'kg', 8, 28, 8.5, 26.00),
    ('Contramuslo pollo', 'kg', None, None, None, 4.20),
    ('Lubina', 'kg', 6, 27, 5, 14.00),
    ('Langostino', 'kg', None, None, None, 18.00),
    ('Patata', 'kg', 25, 30, 22, 1.10),
    ('Tomate', 'kg', 6, 12, 6.5, 2.40),
    ('Lechuga', 'kg', 4, 18, 4.5, 1.80),
    ('Mantequilla', 'kg', 3, 2, 2.5, 9.00),
    ('Aceite oliva', 'L', 10, 5, 11.5, 8.50),
    ('Vino blanco', 'L', 12, 6, 14, 4.50),
    ('Harina', 'kg', None, None, None, 0.90),
    ('Azúcar', 'kg', None, None, None, 1.20),
    ('Huevos', 'docena', None, None, None, 3.60),
    ('Nata', 'L', None, None, None, 2.80),
    ('Chocolate', 'kg', None, None, None, 11.00),
]
# (plato, raciones vendidas, {producto: cantidad BRUTA por ración})
VENTAS = [
    ('Solomillo al Pedro Ximénez', 120,
     {'Solomillo ternera': 0.22, 'Patata': 0.18, 'Mantequilla': 0.02,
      'Vino blanco': 0.03}),
    ('Lubina a la plancha', 80,
     {'Lubina': 0.34, 'Patata': 0.12, 'Aceite oliva': 0.02}),
    ('Ensalada de la casa', 150,
     {'Tomate': 0.07, 'Lechuga': 0.11, 'Aceite oliva': 0.01}),
]
FILAS_VENTAS = 10          # rows 5..14 de «Ventas del periodo»
FILAS_REGISTRO = 90        # DOM-25/COM-32: 90 incidencias al mes ≈ 3 al día

MOTIVOS = ('Caducidad,Mal estado,Sobreproducción,Error de preparación,'
           'Almacenaje incorrecto,Porcionado excesivo,Otro')


# ==========================================================================
# Utilidades
# ==========================================================================
def _reset(ws, hasta_fila, hasta_col):
    """Deja la hoja en blanco absoluto: valores, estilos, merges, validaciones,
    formato condicional y gráficos. Es lo que hace idempotente al módulo."""
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    ws.conditional_formatting = ConditionalFormattingList()
    ws.data_validations.dataValidation = []
    ws._charts = []
    filas = max(hasta_fila, ws.max_row)
    cols = max(hasta_col, ws.max_column)
    for fila in range(1, filas + 1):
        for col in range(1, cols + 1):
            c = ws.cell(row=fila, column=col)
            c.value = None
            c.fill = PatternFill()
            c.font = Font()
            c.number_format = 'General'
            c.alignment = Alignment()
            c.protection = Protection(locked=True)


def _c(ws, fila, col, valor=None, fmt=None, fill=None, verde=False,
       bold=False, italic=False, size=11, color=None, align=None,
       wrap=False, registro=None):
    cel = ws.cell(row=fila, column=col)
    cel.value = valor
    if verde:
        fill = motor.VERDE
    cel.fill = PatternFill('solid', fgColor=fill) if fill else PatternFill()
    cel.protection = Protection(locked=not verde)
    cel.font = Font(bold=bold, italic=italic, size=size, color=color)
    if fmt:
        cel.number_format = fmt
    if align or wrap:
        cel.alignment = Alignment(horizontal=align, vertical='center',
                                  wrap_text=wrap)
    if registro is not None and isinstance(valor, str) and valor.startswith('='):
        registro.append((ws.title, cel.coordinate, valor))
    return cel


def _titulo(ws, fila, col0, col1, texto, size=13):
    _c(ws, fila, col0, texto, bold=True, size=size)
    if col1 > col0:
        ws.merge_cells(start_row=fila, start_column=col0,
                       end_row=fila, end_column=col1)


def _nota(ws, fila, col0, col1, texto):
    _c(ws, fila, col0, texto, italic=True, size=9, color='606060', wrap=True)
    if col1 > col0:
        ws.merge_cells(start_row=fila, start_column=col0,
                       end_row=fila, end_column=col1)
    ws.row_dimensions[fila].height = 26


def _cabecera(ws, fila, col0, textos, anchos):
    for i, t in enumerate(textos):
        c = _c(ws, fila, col0 + i, t, fill=motor.CAB, align='center', wrap=True)
        c.font = Font(bold=True, color='FFFFFF', size=10)
    for i, a in enumerate(anchos):
        ws.column_dimensions[get_column_letter(col0 + i)].width = a
    ws.row_dimensions[fila].height = 30


def _hoja(wb, nombre, despues=None):
    if nombre in wb.sheetnames:
        return wb[nombre]
    ws = wb.create_sheet(nombre)
    if despues and despues in wb.sheetnames:
        wb._sheets.remove(ws)
        wb._sheets.insert(wb._sheets.index(wb[despues]) + 1, ws)
    return ws


def _ocultar_na(ws, rango):
    """Fuente blanca sobre las celdas que devuelven #N/A (TEC-R2-02).

    El #N/A es lo que hace que el gráfico SALTE el punto en vez de dibujar un
    cero; en la tabla no aporta nada, así que se esconde. Es la receta
    estándar. Idempotente: la regla se borra antes de volver a añadirla.
    """
    try:
        reglas = ws.conditional_formatting._cf_rules
    except AttributeError:
        reglas = None
    if reglas is not None:
        for r in list(reglas.keys()):
            if str(r.sqref) == rango:
                del reglas[r]
    primera = rango.split(':')[0]
    ws.conditional_formatting.add(
        rango,
        FormulaRule(formula=[f'ISNA({primera})'],
                    font=Font(color='FFFFFF'), stopIfTrue=True))


def _semaforo(ws, rango):
    """OK verde / ALERTA rojo sobre una columna de Estado."""
    ws.conditional_formatting.add(rango, CellIsRule(
        operator='equal', formula=['"ALERTA"'],
        fill=PatternFill('solid', start_color=motor.ROJO_BG,
                         end_color=motor.ROJO_BG),
        font=Font(color=motor.ROJO_FG, bold=True), stopIfTrue=False))
    ws.conditional_formatting.add(rango, CellIsRule(
        operator='equal', formula=['"OK"'],
        fill=PatternFill('solid', start_color=OK_BG, end_color=OK_BG),
        font=Font(color=OK_FG, bold=True), stopIfTrue=False))


def _cats_texto(ch, ref):
    """openpyxl 3.1.3 `set_categories()` emite SIEMPRE <numRef> (verificado en
    el código de `ChartBase.set_categories`), y sobre celdas de TEXTO eso deja
    el eje sin rótulos — es literalmente el bug COM-24/TEC-13 del gráfico que
    ya venía en el 11. Se fija a mano como <strRef>."""
    for s in ch.series:
        s.cat = AxDataSource(strRef=StrRef(f=str(ref)))


def _ejes(ch, fmt_y=None):
    ch.x_axis.axPos = 'b'          # eje de categorías ABAJO (COM-24)
    ch.y_axis.axPos = 'l'
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    if fmt_y:
        ch.y_axis.numFmt = fmt_y
    ch.dispBlanksAs = 'gap'


# ==========================================================================
# Instrucciones (las del motor son las de una hoja de escandallo)
# ==========================================================================
PIE = motor.PIE
COPY = motor.COPY


def _instrucciones(wb, secciones, informe):
    """`secciones` = [(título|None, [líneas])]. La línea de versión se escribe
    aquí (igual que hace `motor.reescribir_instrucciones`) para que el
    `linea_instrucciones()` de `motor.cerrar` la encuentre y NO añada otra: sin
    esto la 2.ª pasada la duplicaría dos filas más abajo."""
    ws = wb['Instrucciones']
    col = 2 if ws.cell(row=2, column=2).value else 1
    modelo_titulo = copy.copy(ws.cell(row=6, column=col)._style)
    modelo_linea = copy.copy(ws.cell(row=8, column=col)._style)
    for r in range(6, ws.max_row + 1):
        ws.cell(row=r, column=col).value = None

    fila = 6
    for titulo, lineas in secciones:
        if titulo:
            c = ws.cell(row=fila, column=col, value=titulo)
            c._style = copy.copy(modelo_titulo)
            fila += 2
        for t in lineas:
            c = ws.cell(row=fila, column=col, value=t)
            c._style = copy.copy(modelo_linea)
            fila += 1
        fila += 1
    for t in (PIE, None, COPY, None, motor.VERSION_LINE):
        if t is None:
            fila += 1
            continue
        c = ws.cell(row=fila, column=col, value=t)
        c._style = copy.copy(modelo_linea)
        fila += 1
    informe.append('Instrucciones reescritas (§4)')


# ==========================================================================
# 09 · control de mermas
# ==========================================================================
def _f09(wb, informe, registro):
    ws = wb['Mermas Semanal']
    _reset(ws, 24, 10)

    _titulo(ws, 1, 1, 10, 'Control de Mermas — Desperdicio semanal sobre compra')
    _c(ws, 2, 1, 'Semana del:', bold=True)
    _c(ws, 2, 2, None, fmt=FMT_FECHA, verde=True, align='center')

    _cabecera(ws, 4, 1, [
        'Categoría', 'Objetivo mín.', 'Objetivo típico', 'Objetivo máx.',
        'Compra total (€)', 'Desperdicio real (€)', 'Desperdicio real (%)',
        'Diferencia (p.p.)', 'Estado', 'Notas',
    ], [20, 13, 14, 13, 16, 18, 18, 15, 11, 30])

    for i, (cat, mn, tip, mx) in enumerate(DESPERDICIO):
        fila = 5 + i
        banda = BANDA_A if i % 2 == 0 else BANDA_B
        compra, real = EJEMPLO_09.get(cat, (None, None))
        _c(ws, fila, 1, cat, fill=banda)
        _c(ws, fila, 2, mn, fmt=FMT_PCT1, fill=banda, align='center')
        _c(ws, fila, 3, tip, fmt=FMT_PCT1, verde=True, align='center')
        _c(ws, fila, 4, mx, fmt=FMT_PCT1, fill=banda, align='center')
        _c(ws, fila, 5, compra, fmt=FMT_EUR, verde=True)
        _c(ws, fila, 6, real, fmt=FMT_EUR, verde=True)
        _c(ws, fila, 7, f'=IFERROR(IF(E{fila}=0,"",F{fila}/E{fila}),"")',
           fmt=FMT_PCT1, fill=banda, align='center', registro=registro)
        _c(ws, fila, 8, f'=IFERROR(IF(G{fila}="","",G{fila}-C{fila}),"")',
           fmt=FMT_PP, fill=banda, align='center', registro=registro)
        _c(ws, fila, 9,
           f'=IFERROR(IF(G{fila}="","",IF(G{fila}<=C{fila},"OK","ALERTA")),"")',
           fill=banda, align='center', registro=registro)
        _c(ws, fila, 10, None, verde=True)

    tot = 5 + len(DESPERDICIO)
    for col in range(1, 11):
        _c(ws, tot, col, None, fill=motor.ORO, bold=True)
    _c(ws, tot, 1, 'TOTAL', fill=motor.ORO, bold=True)
    _c(ws, tot, 5, f'=SUM(E5:E{tot - 1})', fmt=FMT_EUR, fill=motor.ORO,
       bold=True, registro=registro)
    _c(ws, tot, 6, f'=SUM(F5:F{tot - 1})', fmt=FMT_EUR, fill=motor.ORO,
       bold=True, registro=registro)
    _c(ws, tot, 7, f'=IFERROR(IF(E{tot}=0,"",F{tot}/E{tot}),"")',
       fmt=FMT_PCT1, fill=motor.ORO, bold=True, align='center',
       registro=registro)

    _nota(ws, tot + 2, 1, 10,
          'Aquí se registra el DESPERDICIO (caducidad, mal estado, '
          'sobreproducción, porcionado): euros tirados sobre euros comprados. '
          'La merma de despiece y limpieza NO va aquí — ya está cobrada dentro '
          'del coste del plato, en las plantillas de escandallo (hoja '
          '«Mermas»). Mezclarlas era el error de la v1.1: con un objetivo del '
          '45 % un marisco podrido daba «OK».')
    _nota(ws, tot + 3, 1, 10,
          'Objetivos de referencia sobre la compra: secos y congelados 2-3 % · '
          'carne y pescado 3-5 % · verdura y fruta 5-8 % · bebidas 1-2 % · pan '
          'y bollería 5-10 % (sobreproducción). El «objetivo típico» es la '
          'celda verde: ajústalo a tu casa; mínimo y máximo son la horquilla '
          'del sector.')

    _semaforo(ws, f'I5:I{tot - 1}')
    ws.conditional_formatting.add(f'G5:G{tot - 1}', ColorScaleRule(
        start_type='min', start_color=OK_BG,
        mid_type='percentile', mid_value=50, mid_color=AMBAR,
        end_type='max', end_color=motor.ROJO_BG))
    ws.freeze_panes = 'A5'
    informe.append('Mermas Semanal: columna B→D = desperdicio objetivo '
                   'mín/típico/máx s/compra (DOM-02/TEC-04) + semáforo con '
                   'formato condicional (TEC-03/COM-08) + escala de color')
    return tot


def _f09_evolucion(wb, informe, registro, fila_total):
    ws = _hoja(wb, 'Evolución', despues='Mermas Semanal')
    _reset(ws, 18, 5)
    _titulo(ws, 1, 1, 5, 'Evolución del desperdicio — 12 semanas')
    _nota(ws, 2, 1, 5,
          'La fila «Semana en curso» está ENLAZADA a «Mermas Semanal»: cambia '
          'sola según rellenas la semana. Cada lunes, ANTES de vaciar el '
          'registro semanal, copia su total (compra y desperdicio) en la fila '
          'de la semana que cierras y pega VALORES, no fórmulas — empezando '
          'por la Semana 1, cuyo enlace se sustituye al pegar encima. Si no lo '
          'haces, la semana 1 seguirá enseñando lo que haya en la hoja '
          'semanal y perderás el histórico, que es justo lo que esta pestaña '
          'existe para guardar.')
    _c(ws, 3, 1, 'Objetivo global de desperdicio (%)', bold=True)
    _c(ws, 3, 2, OBJETIVO_EVOL, fmt=FMT_PCT1, verde=True, align='center')

    _cabecera(ws, 5, 1, ['Semana', 'Compra total (€)', 'Desperdicio (€)',
                         'Desperdicio (%)', 'Objetivo (%)'],
              [34, 18, 18, 16, 14])

    for i in range(12):
        fila = 6 + i
        banda = BANDA_A if i % 2 == 0 else BANDA_B
        if i == 0:
            _c(ws, fila, 1,
               'Semana en curso (enlazada a «Mermas Semanal» — pega encima '
               'el valor al cerrarla)', fill=banda)
            _c(ws, fila, 2, f"='Mermas Semanal'!E{fila_total}", fmt=FMT_EUR,
               fill=banda, registro=registro)
            _c(ws, fila, 3, f"='Mermas Semanal'!F{fila_total}", fmt=FMT_EUR,
               fill=banda, registro=registro)
        else:
            ej = EJEMPLO_EVOL[i - 1] if i - 1 < len(EJEMPLO_EVOL) else (None, None)
            _c(ws, fila, 1, f'Semana {i + 1}', fill=banda)
            _c(ws, fila, 2, ej[0], fmt=FMT_EUR, verde=True)
            _c(ws, fila, 3, ej[1], fmt=FMT_EUR, verde=True)
        _c(ws, fila, 4, f'=IF(B{fila}=0,NA(),C{fila}/B{fila})',
           fmt=FMT_PCT1, fill=banda, align='center', registro=registro)
        _c(ws, fila, 5, f'=IF(B{fila}=0,NA(),$B$3)',
           fmt=FMT_PCT1, fill=banda, align='center', registro=registro)

    _ocultar_na(ws, 'D6:E17')
    ch = LineChart()
    ch.title = 'Desperdicio semanal vs objetivo'
    ch.style = 2
    ch.height, ch.width = 9, 18
    datos = Reference(ws, min_col=4, max_col=5, min_row=5, max_row=17)
    ch.add_data(datos, titles_from_data=True)
    _cats_texto(ch, Reference(ws, min_col=1, min_row=6, max_row=17))
    _ejes(ch, FMT_PCT1)
    # TEC-R2-10: anclado en G5 el gráfico caía TRES columnas a la derecha del
    # área de impresión y quien exportaba a PDF se llevaba la tabla sin la
    # gráfica. Estirar el área hasta la R metía 13 columnas vacías y el ajuste
    # a una página lo dejaba al 54 %. Debajo de la tabla ocupa exactamente el
    # ancho A:E (18 cm ≈ 7,3 in) y entra entero en un A4 apaisado.
    ws.add_chart(ch, 'A19')
    ws.freeze_panes = 'A6'
    informe.append('hoja «Evolución» (12 semanas alimentadas por el TOTAL) '
                   'con LineChart de openpyxl (COM-09/DOM-26)')


INS_09 = [
    (None, [
        '▸ Esta plantilla mide el DESPERDICIO: los euros que se tiran '
        '(caducidad, mal estado, sobreproducción, porcionado) sobre los euros '
        'que se compran.',
        '▸ NO mide la merma de despiece y limpieza. Esa va en las plantillas '
        'de escandallo (hoja «Mermas») y ya está cobrada dentro del precio del '
        'plato: registrarla otra vez aquí sería contarla dos veces.',
    ]),
    ('Cómo se rellena', [
        "▸ Pestaña 'Mermas Semanal': escribe la fecha del lunes y, por "
        'categoría, la COMPRA de la semana y los EUROS desperdiciados.',
        '▸ El porcentaje, la diferencia en puntos y el estado salen solos.',
        '▸ Sólo se escriben las celdas VERDES. El resto son fórmulas.',
    ]),
    ('El semáforo', [
        '▸ Estado en VERDE («OK»): el desperdicio está en el objetivo o por '
        'debajo.',
        '▸ Estado en ROJO («ALERTA»): te estás pasando. Empieza por ahí.',
        '▸ La columna «Desperdicio real (%)» lleva además una escala de color: '
        'de un vistazo se ve qué categoría se sale de la fila.',
    ]),
    ('Los objetivos (mín. / típico / máx.)', [
        '▸ Secos y congelados 2-3 % · carne y pescado 3-5 % · verdura y fruta '
        '5-8 % · bebidas 1-2 % · pan y bollería 5-10 %.',
        '▸ El «objetivo típico» es la celda verde y es el que dispara la '
        'alerta: bájalo cuando tu casa mejore.',
        '▸ Mínimo y máximo son la horquilla del sector: sirven para saber si '
        'tu número es normal, bueno o malo.',
    ]),
    ('Pestaña «Evolución»', [
        '▸ Doce semanas. La primera fila («Semana en curso») está ENLAZADA al '
        'TOTAL de «Mermas Semanal»: se mueve sola mientras rellenas la '
        'semana.',
        '▸ Cada lunes, ANTES de vaciar el registro semanal, copia su total en '
        'la fila de la semana que cierras y pega VALORES, no fórmulas. La '
        'primera vez pegarás encima del enlace, y así debe ser: si lo dejas '
        'vivo, la semana 1 pasará a enseñar los datos de la semana 2 y '
        'perderás el histórico.',
        '▸ El gráfico compara tu desperdicio con el objetivo global semana a '
        'semana: la tendencia es lo que decide, no la foto de un lunes.',
        '▸ Las semanas todavía sin rellenar devuelven #N/A a propósito: es el '
        'único valor que el gráfico SALTA. Con una celda vacía o un texto, '
        'Excel dibujaría un cero y parecería que tu desperdicio se desplomó.',
    ]),
    ('Protección de la hoja', [
        '▸ Las hojas están protegidas SIN contraseña para que no borres una '
        'fórmula sin querer: sólo se escriben las celdas verdes.',
        '▸ Para tocar cualquier otra cosa: Revisar → Desproteger hoja. No pide '
        'contraseña.',
    ]),
]


# ==========================================================================
# 10 · calculadora de PVP
# ==========================================================================
def _f10(wb, informe, registro):
    ws = wb['Calculadora PVP']
    _reset(ws, 20, 10)
    ws.column_dimensions['A'].width = 3

    _titulo(ws, 2, 2, 10, 'Calculadora de PVP Sugerido')
    _c(ws, 4, 2, 'COSTE POR RACIÓN / POR UNIDAD (€):', bold=True)
    _c(ws, 4, 3, COSTE_10, fmt=FMT_EUR, verde=True, align='center')
    _c(ws, 5, 2, 'Tipo de IVA (%):', bold=True)
    _c(ws, 5, 3, IVA_10, fmt=motor.FMT_PCT, verde=True, align='center')
    _nota(ws, 6, 2, 10,
          'El IVA es una celda: 10 % en hostelería en España; cámbialo si te '
          'aplica otro tipo (IGIC en Canarias, o el IVA/ITBIS/IVU de tu país). '
          'El «Multiplicador» es 1 ÷ media del rango de food cost: es el '
          'número por el que multiplicas el coste para sacar el PVP.')

    _cabecera(ws, 8, 2, [
        'Tipo de establecimiento', 'Food Cost mín.', 'Food Cost máx.',
        'Multiplicador', 'Comisión plataforma (%)', 'PVP sin IVA (€)',
        'PVP con IVA (€)', 'Ingreso neto tras comisión (€)',
        'Margen neto (€)',
    ], [28, 13, 13, 13, 17, 16, 16, 20, 15])

    for i, (nombre, mn, mx, com) in enumerate(TIPOS_10):
        fila = 9 + i
        banda = BANDA_A if i % 2 == 0 else BANDA_B
        _c(ws, fila, 2, nombre, fill=banda)
        _c(ws, fila, 3, mn, fmt=motor.FMT_PCT, verde=True, align='center')
        _c(ws, fila, 4, mx, fmt=motor.FMT_PCT, verde=True, align='center')
        _c(ws, fila, 5, f'=IFERROR(1/AVERAGE(C{fila},D{fila}),"")',
           fmt=FMT_MULT, fill=banda, align='center', registro=registro)
        _c(ws, fila, 6, com, fmt=motor.FMT_PCT, verde=True, align='center')
        _c(ws, fila, 7, f'=IFERROR($C$4*E{fila}/(1-F{fila}),"")', fmt=FMT_EUR,
           fill=banda, bold=True, registro=registro)
        # TEC-R2-11: el crema de énfasis se pintaba sólo en las filas impares
        # y el banding gris se lo comía en las pares: la columna que el usuario
        # mira salía a rayas crema/gris, como un fallo de renderizado. El
        # énfasis gana en TODA la columna.
        _c(ws, fila, 8, f'=IFERROR(G{fila}*(1+$C$5),"")', fmt=FMT_EUR,
           fill=motor.CREMA, registro=registro)
        _c(ws, fila, 9, f'=IFERROR(G{fila}*(1-F{fila}),"")', fmt=FMT_EUR,
           fill=banda, registro=registro)
        _c(ws, fila, 10, f'=IFERROR(I{fila}-$C$4,"")', fmt=FMT_EUR,
           fill=banda, registro=registro)

    ultima = 9 + len(TIPOS_10) - 1
    _nota(ws, ultima + 2, 2, 10,
          'Delivery: la comisión de la plataforma (25-35 %) se descuenta ANTES '
          'de calcular el precio, así que el PVP de la carta de delivery sale '
          'muy por encima del de sala — es la única forma de que te quede el '
          'mismo dinero en el bolsillo. Su food cost objetivo (28-32 %) se '
          'mide sobre el INGRESO NETO, no sobre el precio de la carta: un '
          '30 % sobre neto equivale a un ~21 % sobre el PVP bruto, y por eso '
          'el precio de delivery sube. Si vendes por tu propia web, pon la '
          'comisión a 0 %.')
    _nota(ws, ultima + 3, 2, 10,
          'La comisión funciona en TODAS las filas: si trabajas una carta de '
          'sala también en plataforma, copia el food cost de tu fila y pon la '
          'comisión que te cobren.')
    ws.freeze_panes = 'B9'
    informe.append('Calculadora PVP: fila «Delivery» con comisión descontada '
                   'antes del PVP + columna «Multiplicador» + IVA en celda '
                   '(DOM-11/TEC-18/COM-11)')


INS_10 = [
    (None, [
        '▸ Escribe en la celda verde (C4) el coste de UNA ración. Sale de la '
        'plantilla de escandallo: es la fila «COSTE POR RACIÓN» en 01-04 y '
        '06-08, y «COSTE POR UNIDAD» en 05-pastelería. NO el «COSTE TOTAL DE '
        'LA TANDA»: en pastelería difieren por 12, 20 o 30.',
        '▸ La tabla calcula el PVP para cada tipo de establecimiento. Sólo se '
        'escriben las celdas VERDES.',
    ]),
    ('Qué es cada columna', [
        '▸ Food Cost mín. / máx.: la horquilla de food cost objetivo de ese '
        'tipo de negocio. Editable.',
        '▸ Multiplicador: 1 ÷ media de la horquilla. Es el número por el que '
        'multiplicas el coste para sacar el PVP — el atajo que se usa en '
        'cocina.',
        '▸ Comisión plataforma (%): lo que se queda el intermediario. Se '
        'descuenta ANTES de calcular el precio.',
        '▸ PVP sin IVA: coste × multiplicador ÷ (1 − comisión).',
        '▸ PVP con IVA: el anterior por el tipo de IVA de C5.',
        '▸ Ingreso neto tras comisión: lo que te ingresa la plataforma.',
        '▸ Margen neto: ingreso neto − coste del plato. Es el dinero que te '
        'queda de verdad.',
    ]),
    ('Delivery', [
        '▸ Es la fila que más se equivoca todo el mundo. Aplicar el PVP de '
        '«Casual Dining» a una carta de Glovo o Uber Eats significa regalar la '
        'comisión: con un 30 % de comisión, un plato de 18 € deja 12,60 €.',
        '▸ Su food cost objetivo (28-32 %) se mide sobre el ingreso NETO, ya '
        'descontada la comisión. Un 30 % sobre neto equivale a un ~21 % sobre '
        'el PVP bruto: por eso el precio de la carta de delivery sale alto.',
        '▸ Si vendes por tu propia web o con reparto propio, pon la comisión a '
        '0 % y usa el food cost de tu tipo de local.',
    ]),
    ('Tipo de IVA', [
        '▸ C5 es editable: 10 % en hostelería en España. Cámbialo si te aplica '
        'otro tipo (IGIC en Canarias, o el IVA/ITBIS/IVU de tu país en '
        'Latinoamérica).',
    ]),
    ('Protección de la hoja', [
        '▸ La hoja está protegida SIN contraseña: sólo se escriben las celdas '
        'verdes. Revisar → Desproteger hoja para tocar el resto.',
    ]),
]


# ==========================================================================
# 11 · dashboard de food cost mensual
# ==========================================================================
def _f11(wb, informe, registro):
    ws = wb['Dashboard']
    _reset(ws, 20, 11)
    ws.column_dimensions['A'].width = 3

    _titulo(ws, 2, 2, 11, 'Dashboard Food Cost Mensual')
    # TEC-27: «Food Cost objetivo (%):» son 23 caracteres y la columna B es la
    # de los meses (14). El rótulo se combina en B:C y el input baja a D, que
    # es la solución que no obliga a dejar la columna de meses desproporcionada.
    _c(ws, 3, 2, 'Año:', bold=True)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
    _c(ws, 3, 4, 2026, fmt='0', verde=True, align='center')
    _c(ws, 4, 2, 'Food Cost objetivo (%):', bold=True)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    _c(ws, 4, 4, OBJETIVO_11, fmt=motor.FMT_PCT, verde=True, align='center')
    _nota(ws, 5, 2, 11,
          'Food cost = (stock inicial + compras − stock final) ÷ ventas. NO es '
          'compras ÷ ventas: una compra fuerte a fin de mes te movería el '
          'indicador varios puntos sin que hubiera cambiado nada en la cocina. '
          'Todo SIN IVA: si tu TPV te da las ventas con IVA, divídelas entre '
          '1,10 antes de escribirlas.')

    _cabecera(ws, 6, 2, [
        'Mes', 'Stock inicial (€)', 'Compras netas sin IVA (€)',
        'Stock final (€)', 'Consumo (€)', 'Ventas netas sin IVA (€)',
        'Food Cost %', 'Objetivo %', 'Diferencia', 'Estado',
    ], [14, 15, 19, 15, 15, 19, 13, 12, 13, 11])

    for i, mes in enumerate(MESES):
        fila = 7 + i
        banda = BANDA_A if i % 2 == 0 else BANDA_B
        ini, com, fin, ven = EJEMPLO_11.get(mes, (None, None, None, None))
        _c(ws, fila, 2, mes, fill=banda)
        _c(ws, fila, 3, ini, fmt=FMT_EUR0, verde=True)
        _c(ws, fila, 4, com, fmt=FMT_EUR0, verde=True)
        _c(ws, fila, 5, fin, fmt=FMT_EUR0, verde=True)
        _c(ws, fila, 6,
           f'=IFERROR(IF(G{fila}=0,"",C{fila}+D{fila}-E{fila}),"")',
           fmt=FMT_EUR0, fill=banda, registro=registro)
        _c(ws, fila, 7, ven, fmt=FMT_EUR0, verde=True)
        _c(ws, fila, 8, f'=IF(G{fila}=0,NA(),F{fila}/G{fila})',
           fmt=FMT_PCT1, fill=banda, bold=True, align='center',
           registro=registro)
        _c(ws, fila, 9, f'=IF(G{fila}=0,NA(),$D$4)', fmt=FMT_PCT1,
           fill=banda, align='center', registro=registro)
        _c(ws, fila, 10, f'=IFERROR(H{fila}-I{fila},"")',
           fmt=FMT_PP, fill=banda, align='center', registro=registro)
        _c(ws, fila, 11,
           f'=IFERROR(IF(H{fila}<=$D$4,"OK","ALERTA"),"")',
           fill=banda, align='center', registro=registro)

    tot = 7 + len(MESES)
    for col in range(2, 12):
        _c(ws, tot, col, None, fill=motor.ORO, bold=True)
    _c(ws, tot, 2, 'TOTAL ANUAL', fill=motor.ORO, bold=True)
    _c(ws, tot, 4, f'=SUM(D7:D{tot - 1})', fmt=FMT_EUR0, fill=motor.ORO,
       bold=True, registro=registro)
    _c(ws, tot, 6, f'=SUM(F7:F{tot - 1})', fmt=FMT_EUR0, fill=motor.ORO,
       bold=True, registro=registro)
    _c(ws, tot, 7, f'=SUM(G7:G{tot - 1})', fmt=FMT_EUR0, fill=motor.ORO,
       bold=True, registro=registro)
    _c(ws, tot, 8, f'=IFERROR(IF(G{tot}=0,"",F{tot}/G{tot}),"")',
       fmt=FMT_PCT1, fill=motor.ORO, bold=True, align='center',
       registro=registro)
    _c(ws, tot, 11,
       f'=IFERROR(IF(H{tot}="","",IF(H{tot}<=$D$4,"OK","ALERTA")),"")',
       fill=motor.ORO, bold=True, align='center', registro=registro)
    _ocultar_na(ws, f'H7:I{tot - 1}')

    _semaforo(ws, f'K7:K{tot}')
    ws.conditional_formatting.add(f'H7:H{tot}', FormulaRule(
        formula=[f'AND(ISNUMBER(H7),H7>$D$4)'],
        fill=PatternFill('solid', start_color=motor.ROJO_BG,
                         end_color=motor.ROJO_BG),
        font=Font(color=motor.ROJO_FG, bold=True), stopIfTrue=False))

    # ---- gráfico REGENERADO (COM-24/TEC-13/TEC-27) ----------------------
    ch = BarChart()
    ch.type = 'col'
    ch.title = 'Food Cost mensual vs objetivo'
    ch.style = 2
    ch.height, ch.width = 9, 18          # 3,5 × 7,1 in: cabe en un A4 apaisado
    datos = Reference(ws, min_col=8, max_col=9, min_row=6, max_row=tot - 1)
    ch.add_data(datos, titles_from_data=True)
    _cats_texto(ch, Reference(ws, min_col=2, min_row=7, max_row=tot - 1))
    _ejes(ch, FMT_PCT1)
    ws.add_chart(ch, 'B21')
    ws.freeze_panes = 'B7'
    informe.append('Dashboard: stock inicial/final → food cost = '
                   'consumo/ventas (DOM-03), rótulos «sin IVA» (TEC-14), '
                   'columna Estado + formato condicional (COM-10), serie '
                   'objetivo condicional (TEC-13) y gráfico regenerado con '
                   'categorías de texto y eje abajo (COM-24)')


INS_11 = [
    (None, [
        '▸ Registra, mes a mes, el stock inicial, las compras, el stock final '
        'y las ventas. Sólo se escriben las celdas VERDES.',
        '▸ El food cost, la diferencia con tu objetivo, el estado y el gráfico '
        'salen solos.',
    ]),
    ('La fórmula (y por qué importa)', [
        '▸ Consumo = stock inicial + compras − stock final.',
        '▸ Food cost = consumo ÷ ventas.',
        '▸ NO es compras ÷ ventas. Una compra fuerte el día 28, o un vaciado '
        'de cámara, te movería el indicador varios puntos sin que hubiera '
        'cambiado nada en la cocina — que es justo el error que este dashboard '
        'existe para evitar.',
        '▸ El stock en euros lo sacas de la pestaña «Inventario» del BONUS: '
        'cantidad × precio unitario.',
    ]),
    ('IVA: todo en NETO', [
        '▸ Compras y ventas van SIN IVA. Si tu TPV te da las ventas con IVA '
        'incluido, divídelas entre 1,10 (o entre 1 + tu tipo) antes de '
        'escribirlas.',
        '▸ Con las ventas en bruto el food cost sale unos 3 puntos mejor de lo '
        'real y te crees que vas bien.',
    ]),
    ('Alertas', [
        '▸ Columna «Estado»: VERDE («OK») si el food cost del mes está en el '
        'objetivo o por debajo; ROJO («ALERTA») si se pasa.',
        '▸ La columna «Food Cost %» se pinta de rojo cuando supera el '
        'objetivo.',
        '▸ El gráfico compara cada mes con el objetivo. Los meses sin datos no '
        'pintan barra: ni la real ni la del objetivo — sus celdas devuelven '
        '#N/A a propósito (es el único valor que el gráfico salta) y el '
        'formato condicional lo esconde de la tabla.',
    ]),
    ('Protección de la hoja', [
        '▸ La hoja está protegida SIN contraseña: sólo se escriben las celdas '
        'verdes. Revisar → Desproteger hoja para tocar el resto.',
    ]),
]


# ==========================================================================
# BONUS · inventario + registro de incidencias
# ==========================================================================
def _bonus_ventas(wb, informe, registro):
    """Pestaña «Ventas del periodo»: plato × raciones × cantidad/ración.
    La fila de TOTALES es la que alimenta el consumo teórico del inventario."""
    ws = _hoja(wb, 'Ventas del periodo', despues='Inventario')
    n = len(INVENTARIO)
    ultima_col = 2 + n                      # A plato, B raciones, C..Q productos
    _reset(ws, 18, ultima_col)

    _titulo(ws, 1, 1, ultima_col, 'Ventas del periodo — raciones vendidas por plato')
    _nota(ws, 2, 1, ultima_col,
          'Una fila por plato: cuántas raciones vendiste y cuánto lleva CADA '
          'ración de cada producto del inventario. La cantidad es la BRUTA del '
          'escandallo (columna «Cant. Bruta», ya con la merma dentro). La fila '
          'de TOTALES multiplica raciones × cantidad y es el «Consumo teórico» '
          'que aparece en la pestaña «Inventario».')

    _c(ws, 3, 2, 'Unidad →', bold=True, size=9, align='right')
    for i in range(n):
        _c(ws, 3, 3 + i, f'=Inventario!B{5 + i}', size=9, color='808080',
           align='center', registro=registro)

    _cabecera(ws, 4, 1,
              ['Plato', 'Raciones vendidas'] + [f'=Inventario!A{5 + i}'
                                                for i in range(n)],
              [28, 14] + [13] * n)
    for i in range(n):
        registro.append((ws.title, f'{get_column_letter(3 + i)}4',
                         f'=Inventario!A{5 + i}'))

    idx = {p[0]: i for i, p in enumerate(INVENTARIO)}
    for i in range(FILAS_VENTAS):
        fila = 5 + i
        plato, raciones, receta = (VENTAS[i] if i < len(VENTAS)
                                   else (None, None, {}))
        _c(ws, fila, 1, plato, verde=True)
        _c(ws, fila, 2, raciones, fmt='0', verde=True, align='center')
        for j in range(n):
            prod = INVENTARIO[j][0]
            _c(ws, fila, 3 + j, receta.get(prod), fmt='0.000', verde=True,
               align='center')

    tot = 5 + FILAS_VENTAS
    for col in range(1, ultima_col + 1):
        _c(ws, tot, col, None, fill=motor.ORO, bold=True)
    _c(ws, tot, 1, 'TOTALES (raciones · consumo teórico)', fill=motor.ORO,
       bold=True)
    _c(ws, tot, 2, f'=SUM(B5:B{tot - 1})', fmt='0', fill=motor.ORO, bold=True,
       align='center', registro=registro)
    for j in range(n):
        letra = get_column_letter(3 + j)
        _c(ws, tot, 3 + j,
           f'=IFERROR(SUMPRODUCT($B$5:$B${tot - 1},{letra}5:{letra}{tot - 1}),"")',
           fmt=FMT_UD, fill=motor.ORO, bold=True, align='center',
           registro=registro)

    _nota(ws, tot + 2, 1, ultima_col,
          'Los nombres y las unidades de la cabecera se leen de la pestaña '
          '«Inventario»: si cambias un producto allí, aquí cambia solo.')
    ws.freeze_panes = 'C5'
    informe.append(f'hoja «Ventas del periodo» ({FILAS_VENTAS} platos × {n} '
                   'productos) con el consumo teórico por SUMPRODUCT '
                   '(DOM-15/TEC-19)')
    return tot


def _bonus_inventario(wb, informe, registro, fila_totales_ventas):
    ws = wb['Inventario']
    n = len(INVENTARIO)
    _reset(ws, 22, 10)

    _titulo(ws, 1, 1, 10, 'Control de Inventario — Semanal/Mensual')
    _c(ws, 2, 1, 'Periodo:', bold=True)
    _c(ws, 2, 2, None, verde=True, align='center')
    _nota(ws, 3, 1, 10,
          'Consumo real = stock inicial + compras − stock final. Consumo '
          'teórico = lo que DEBERÍAS haber gastado según lo que vendiste '
          '(pestaña «Ventas del periodo»). La diferencia, valorada en euros, '
          'es la pérdida oculta: roturas, robos, porcionado descontrolado o '
          'escandallos desactualizados.')

    _cabecera(ws, 4, 1, [
        'Producto', 'Ud.', 'Stock inicial', 'Compras', 'Stock final',
        'Consumo real', 'Consumo teórico', 'Diferencia (ud.)',
        'Precio unitario (€)', 'Valor de la diferencia (€)',
    ], [22, 10, 13, 12, 13, 14, 15, 15, 17, 21])

    for i, (prod, ud, ini, com, fin, precio) in enumerate(INVENTARIO):
        fila = 5 + i
        banda = BANDA_A if i % 2 == 0 else BANDA_B
        letra = get_column_letter(3 + i)
        # TEC-R2-04: «Producto» y «Ud.» son datos precargados, no rótulos de
        # estructura: la lista de productos es lo PRIMERO que cada restaurante
        # cambia. Bloqueadas dejaban al cliente rellenando stocks de los 15
        # productos del ejemplo (solomillo, lubina, langostino, prosecco…), y
        # con ellas se congelaba también la matriz de «Ventas del periodo»,
        # que lee sus cabeceras de estas mismas celdas.
        _c(ws, fila, 1, prod, verde=True)
        _c(ws, fila, 2, ud, verde=True, align='center')
        _c(ws, fila, 3, ini, fmt=FMT_UD, verde=True, align='center')
        _c(ws, fila, 4, com, fmt=FMT_UD, verde=True, align='center')
        _c(ws, fila, 5, fin, fmt=FMT_UD, verde=True, align='center')
        _c(ws, fila, 6,
           f'=IFERROR(IF(SUM(C{fila}:E{fila})=0,"",C{fila}+D{fila}-E{fila}),"")',
           fmt=FMT_UD, fill=banda, align='center', registro=registro)
        _c(ws, fila, 7,
           f"=IFERROR(IF('Ventas del periodo'!{letra}{fila_totales_ventas}=0,"
           f'"",\'Ventas del periodo\'!{letra}{fila_totales_ventas}),"")',
           fmt=FMT_UD, fill=banda, align='center', registro=registro)
        _c(ws, fila, 8,
           f'=IFERROR(IF(F{fila}="","",IF(G{fila}="","",F{fila}-G{fila})),"")',
           fmt=FMT_DUD, fill=banda, align='center', registro=registro)
        _c(ws, fila, 9, precio, fmt=FMT_EUR, verde=True, align='center')
        _c(ws, fila, 10, f'=IFERROR(IF(H{fila}="","",H{fila}*I{fila}),"")',
           fmt=FMT_EUR, fill=banda, bold=True, registro=registro)

    tot = 5 + n
    for col in range(1, 11):
        _c(ws, tot, col, None, fill=motor.ORO, bold=True)
    _c(ws, tot, 1, 'TOTAL DESVÍO VALORADO', fill=motor.ORO, bold=True)
    ws.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=9)
    _c(ws, tot, 10, f'=SUM(J5:J{tot - 1})', fmt=FMT_EUR, fill=motor.ORO,
       bold=True, registro=registro)

    _nota(ws, tot + 2, 1, 10,
          'Un desvío POSITIVO significa que has gastado más de lo que vendiste '
          'justifica. Empieza siempre por el producto con más euros de desvío, '
          'no por el que más unidades se desvía.')
    _nota(ws, tot + 3, 1, 10,
          'Los productos sin ventas declaradas dejan el consumo teórico y la '
          'diferencia en blanco: sin ventas no hay teórico con el que '
          'comparar, y un cero ahí haría parecer pérdida todo el consumo.')
    ws.freeze_panes = 'A5'
    informe.append('Inventario: consumo teórico enlazado a «Ventas del '
                   'periodo», precio unitario y valor de la diferencia, con '
                   'guarda de vacío (COM-31/TEC-28)')


def _bonus_registro(wb, informe, registro):
    ws = wb['Checklist Mermas']
    _reset(ws, FILAS_REGISTRO + 6, 6)

    _titulo(ws, 1, 1, 6, 'Registro de incidencias (mes)')
    _nota(ws, 2, 1, 6,
          f'{FILAS_REGISTRO} líneas: unas tres incidencias al día durante un '
          'mes. Anota cada merma en cuanto se produce — al final del turno ya '
          'no se acuerda nadie.')
    _cabecera(ws, 3, 1, ['Fecha', 'Producto', 'Cantidad', 'Motivo',
                         'Acción correctiva', 'Coste estimado (€)'],
              [14, 22, 12, 20, 32, 16])

    for i in range(FILAS_REGISTRO):
        fila = 4 + i
        _c(ws, fila, 1, None, fmt=FMT_FECHA, verde=True, align='center')
        _c(ws, fila, 2, None, verde=True)
        _c(ws, fila, 3, None, fmt=FMT_UD, verde=True, align='center')
        _c(ws, fila, 4, None, verde=True)
        _c(ws, fila, 5, None, verde=True)
        _c(ws, fila, 6, None, fmt=FMT_EUR, verde=True)

    tot = 4 + FILAS_REGISTRO
    for col in range(1, 7):
        _c(ws, tot, col, None, fill=motor.ORO, bold=True)
    _c(ws, tot, 1, 'TOTAL MERMAS DEL PERIODO', fill=motor.ORO, bold=True)
    ws.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=5)
    _c(ws, tot, 6, f'=SUM(F4:F{tot - 1})', fmt=FMT_EUR, fill=motor.ORO,
       bold=True, registro=registro)

    dv = DataValidation(type='list', formula1=f'"{MOTIVOS}"', allow_blank=True,
                        showDropDown=False)
    dv.showErrorMessage = True
    dv.errorTitle = 'Valor no válido'
    dv.error = 'Elige un motivo de la lista.'
    ws.add_data_validation(dv)
    dv.add(f'D4:D{tot - 1}')

    ws.freeze_panes = 'A4'
    informe.append(f'Checklist Mermas: «Registro de incidencias (mes)» con '
                   f'{FILAS_REGISTRO} filas, fecha con formato y desplegable '
                   'de motivo extendido (COM-32/DOM-25/TEC-28)')


INS_BONUS = [
    (None, [
        "▸ Pestaña 'Ventas del periodo': lo que has vendido y cuánto lleva "
        'cada ración de cada producto.',
        "▸ Pestaña 'Inventario': el stock real. Compara lo que gastaste con lo "
        'que deberías haber gastado y lo valora en euros.',
        "▸ Pestaña 'Checklist Mermas': el registro de incidencias del mes.",
        '▸ Sólo se escriben las celdas VERDES.',
    ]),
    ('Cómo se calcula el consumo teórico', [
        '▸ Consumo teórico = Σ (raciones vendidas del plato × cantidad de ese '
        'producto por ración).',
        "▸ Se rellena en 'Ventas del periodo': una fila por plato, las "
        'raciones vendidas en la columna B y la cantidad por ración debajo de '
        'cada producto.',
        '▸ La cantidad es la BRUTA del escandallo (columna «Cant. Bruta», con '
        'la merma de despiece ya dentro): si pones la neta, todo el despiece '
        'te saldrá como pérdida oculta.',
        "▸ La fila de TOTALES es la que lee la columna 'Consumo teórico' del "
        'inventario. No hay que copiar nada a mano.',
    ]),
    ('Cómo se lee la diferencia', [
        '▸ Consumo real = stock inicial + compras − stock final.',
        '▸ Diferencia = consumo real − consumo teórico. En positivo, has '
        'gastado de más.',
        '▸ «Valor de la diferencia (€)» = diferencia × precio unitario. Es la '
        'columna que manda: prioriza por euros, no por kilos.',
        '▸ Causas habituales de un desvío positivo: porcionado sin control, '
        'roturas y caídas, robo, o un escandallo desactualizado.',
        '▸ Los productos sin ventas declaradas quedan en blanco a propósito: '
        'sin teórico no hay comparación posible.',
    ]),
    ('Registro de incidencias', [
        '▸ 90 líneas al mes (unas tres al día). Fecha, producto, cantidad, '
        'motivo (desplegable), acción correctiva y coste estimado.',
        '▸ El total del periodo se suma solo al pie de la tabla.',
        '▸ Cruza este total con la pestaña «Mermas Semanal» de '
        '09-control-mermas: si no cuadran, es que hay mermas que no se están '
        'anotando.',
    ]),
    ('Protección de las hojas', [
        '▸ Las hojas están protegidas SIN contraseña: sólo se escriben las '
        'celdas verdes. Revisar → Desproteger hoja para tocar el resto.',
    ]),
]


# ==========================================================================
# API que consume main.py
# ==========================================================================
def pre(wb, fname, informe):
    """Nada que preparar: `post()` reconstruye las hojas enteras."""
    return


# print_area por hoja: `motor.cerrar` fuerza un mínimo de 10 columnas (que es
# lo correcto para una hoja de escandallo, de A a J) y aquí hay hojas de 5 y de
# 17. Se corrige después de cerrar.
# TEC-R2-10: las dos hojas con gráfico tenían el dibujo FUERA de su área de
# impresión — quien exportaba el dashboard o la evolución a PDF se llevaba la
# tabla y ninguna gráfica, que es la única pieza visual del kit y un argumento
# explícito de la landing. El área incluye ahora el ancla del gráfico y las dos
# van en apaisado (`motor.cerrar` pone «Evolución» en vertical porque sólo
# tiene 5 columnas de datos). Valor: (área, apaisado o None para no tocar).
AREAS = {
    F09: {'Mermas Semanal': 'A1:J24',
          # gráfico anclado en A19, 18×9 cm ≈ 12 filas de alto
          'Evolución': ('A1:E36', True, True)},
    F10: {'Calculadora PVP': 'A1:J21'},
    # gráfico anclado en B21, 18×9 cm: acaba dentro de la G y de la fila 38
    F11: {'Dashboard': ('A1:K38', True, True)},
    FBO: {'Inventario': 'A1:J23',
          'Ventas del periodo': 'A1:Q17',
          'Checklist Mermas': f'A1:F{4 + FILAS_REGISTRO}'},
}


def post(wb, fname, informe, registro):
    if fname == F09:
        fila_total = _f09(wb, informe, registro)
        _f09_evolucion(wb, informe, registro, fila_total)
        _instrucciones(wb, INS_09, informe)
    elif fname == F10:
        _f10(wb, informe, registro)
        _instrucciones(wb, INS_10, informe)
    elif fname == F11:
        _f11(wb, informe, registro)
        _instrucciones(wb, INS_11, informe)
    elif fname == FBO:
        fila_tot = _bonus_ventas(wb, informe, registro)
        _bonus_inventario(wb, informe, registro, fila_tot)
        _bonus_registro(wb, informe, registro)
        _instrucciones(wb, INS_BONUS, informe)
    else:
        return

    motor.cerrar(wb, fname, informe)
    for hoja, spec in AREAS.get(fname, {}).items():
        if hoja not in wb.sheetnames:
            continue
        area, apaisado, una_pagina = (spec if isinstance(spec, tuple)
                                      else (spec, None, None))
        wb[hoja].print_area = area
        if apaisado is not None:
            wb[hoja].page_setup.orientation = ('landscape' if apaisado
                                               else 'portrait')
        if una_pagina:
            # tabla + gráfico en UNA hoja: sin esto el salto de página cae en
            # mitad del dibujo y Excel lo manda entero a la página siguiente.
            wb[hoja].page_setup.fitToHeight = 1
    _verificar_graficos(wb, fname, informe)


GRAFICOS_FUERA = []


def _verificar_graficos(wb, fname, informe):
    """Aserción del pipeline (TEC-R2-10): por cada hoja con gráfico, el ancla
    del dibujo tiene que caer DENTRO del área de impresión. Sin esto, reducir
    el gráfico o mover el print_area en pasadas distintas vuelve a sacarlo del
    papel sin que nada avise."""
    for ws in wb.worksheets:
        for ch in getattr(ws, '_charts', []):
            anc = getattr(ch, 'anchor', None)
            if isinstance(anc, str):
                # openpyxl guarda el ancla como cadena ('B21') hasta que el
                # escritor la convierte en OneCellAnchor: en memoria, justo
                # después de add_chart, es un str.
                fil, col = coordinate_to_tuple(anc)
            else:
                desde = getattr(anc, '_from', None) if anc is not None else None
                if desde is None:
                    continue
                col, fil = desde.col + 1, desde.row + 1      # 0-based → 1-based
            area = ws.print_area
            ref = area[0] if isinstance(area, (list, tuple)) else area
            if not ref:
                informe.append(f'AVISO {fname}:{ws.title}: gráfico sin '
                               'print_area')
                continue
            rango = str(ref).split('!')[-1].replace('$', '')
            ini, fin = rango.split(':')
            c1, f1 = coordinate_to_tuple(ini)[1], coordinate_to_tuple(ini)[0]
            c2, f2 = coordinate_to_tuple(fin)[1], coordinate_to_tuple(fin)[0]
            dentro = c1 <= col <= c2 and f1 <= fil <= f2
            msg = (f'{fname}:{ws.title}: gráfico anclado en '
                   f'{get_column_letter(col)}{fil} '
                   + ('DENTRO' if dentro else 'FUERA')
                   + f' del área de impresión {rango}')
            informe.append(msg)
            if not dentro:
                GRAFICOS_FUERA.append(msg)


# ==========================================================================
# Demostraciones (las llama main.py en el paso 7)
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    import contextlib
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}'


def _grafico(path, hoja):
    """Recarga el fichero YA cacheado y describe el gráfico: si openpyxl no
    pudiera releerlo, Excel tampoco lo abriría."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb[hoja]
    fuera = {'hoja': hoja, 'graficos': len(ws._charts), 'series': []}
    for ch in ws._charts:
        fuera['tipo'] = type(ch).__name__
        fuera['x_axis_axPos'] = ch.x_axis.axPos
        fuera['y_axis_axPos'] = ch.y_axis.axPos
        fuera['numFmt_y'] = getattr(ch.y_axis.numFmt, 'formatCode',
                                    ch.y_axis.numFmt)
        fuera['dispBlanksAs'] = getattr(ch, 'dispBlanksAs', None)
        ext = getattr(ch.anchor, 'ext', None)
        # TEC-27: el gráfico de v1.1 medía 9.000.000 EMU (9,84 in) y no cabía
        # en el ancho imprimible de un A4. 6.480.000 EMU = 18 cm = 7,09 in.
        fuera['ext_emu'] = [ext.cx, ext.cy] if ext is not None else None
        for s in ch.series:
            fuera['series'].append({
                'valores': s.val.numRef.f if s.val and s.val.numRef else None,
                'categorias_strRef': (s.cat.strRef.f if s.cat and s.cat.strRef
                                      else None),
                'categorias_numRef': (s.cat.numRef.f if s.cat and s.cat.numRef
                                      else None),
            })
    return fuera


def _celdas(path, hoja, coords, data_only=True):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=data_only)
    ws = wb[hoja]
    return {c: ws[c].value for c in coords}


def demos(carpeta, origen, dir_demos):
    """Devuelve el bloque de demostraciones del grupo C + la lista de fallos."""
    import openpyxl
    fallos = sorted(set(GRAFICOS_FUERA))      # TEC-R2-10
    os.makedirs(dir_demos, exist_ok=True)
    fuera = {}

    # ---- 09 -------------------------------------------------------------
    p09 = os.path.join(carpeta, F09)
    tot = 5 + len(DESPERDICIO)
    wb = openpyxl.load_workbook(p09)
    ws = wb['Mermas Semanal']
    cf = {str(r.sqref): [type(x).__name__ for x in r.rules]
          for r in ws.conditional_formatting}
    filas = []
    wv = openpyxl.load_workbook(p09, data_only=True)['Mermas Semanal']
    for i, (cat, mn, tip, mx) in enumerate(DESPERDICIO):
        f = 5 + i
        if ws.cell(row=f, column=5).value is None:
            continue
        filas.append({
            'ref': f'{F09}:Mermas Semanal:{f}', 'categoria': cat,
            'objetivo_min_tip_max': [mn, tip, mx],
            'compra_eur': ws.cell(row=f, column=5).value,
            'desperdicio_eur': ws.cell(row=f, column=6).value,
            'desperdicio_pct': wv.cell(row=f, column=7).value,
            'diferencia_pp': wv.cell(row=f, column=8).value,
            'estado': wv.cell(row=f, column=9).value,
        })
    # prueba viva sobre una copia desechable: una fila vacía que se pasa
    dst = os.path.join(dir_demos, 'demo-09-alerta.xlsx')
    shutil.copy2(p09, dst)
    wd = openpyxl.load_workbook(dst)
    wsd = wd['Mermas Semanal']
    wsd.protection.sheet = False
    fila_marisco = 5 + [c[0] for c in DESPERDICIO].index('Marisco')
    wsd.cell(row=fila_marisco, column=5, value=500)
    wsd.cell(row=fila_marisco, column=6, value=60)     # 12 % > máx 6 %
    wd.save(dst)
    xl = _pycel(dst)
    viva = {
        'copia_desechable': dst,
        'ref': f'{F09}:Mermas Semanal:{fila_marisco} (Marisco)',
        'se_escribe': {'E': 500, 'F': 60},
        'objetivo_tipico': _ev(xl, f"'Mermas Semanal'!C{fila_marisco}"),
        'desperdicio_pct': _ev(xl, f"'Mermas Semanal'!G{fila_marisco}"),
        'estado': _ev(xl, f"'Mermas Semanal'!I{fila_marisco}"),
    }
    if viva['estado'] != 'ALERTA':
        fallos.append(f'09: desperdicio 12 % sobre un objetivo del 5 % no da '
                      f'ALERTA (da {viva["estado"]!r})')
    ok_ref = 5 + [c[0] for c in DESPERDICIO].index('Carne roja')
    estado_ok = wv.cell(row=ok_ref, column=9).value
    if estado_ok != 'OK':
        fallos.append(f'09: Carne roja al 4,0 % con objetivo 4,0 % debería dar '
                      f'OK y da {estado_ok!r}')
    if f'I5:I{tot - 1}' not in cf:
        fallos.append('09: no hay formato condicional sobre la columna Estado')
    if f'G5:G{tot - 1}' not in cf:
        fallos.append('09: no hay escala de color sobre el desperdicio real')
    graf09 = _grafico(p09, 'Evolución')
    if graf09['graficos'] != 1:
        fallos.append(f'09: la hoja Evolución tiene {graf09["graficos"]} '
                      'gráficos (debería tener 1)')
    for s in graf09['series']:
        if not s['categorias_strRef'] or s['categorias_numRef']:
            fallos.append('09: el LineChart no tiene las categorías como texto')
    fuera['mermas_09'] = {
        'columnas': [ws.cell(row=4, column=c).value for c in range(1, 11)],
        'formato_condicional': cf,
        'filas_con_ejemplo': filas,
        'total': {'compra_eur': wv.cell(row=tot, column=5).value,
                  'desperdicio_eur': wv.cell(row=tot, column=6).value,
                  'desperdicio_pct': wv.cell(row=tot, column=7).value},
        'prueba_viva_alerta': viva,
        'evolucion': {
            'grafico': graf09,
            'semana_1_enlazada': openpyxl.load_workbook(p09)['Evolución']['B6'].value,
            'valores': _celdas(p09, 'Evolución',
                               [f'{c}{r}' for r in range(6, 18)
                                for c in ('B', 'C', 'D', 'E')]),
        },
    }

    # ---- 10 -------------------------------------------------------------
    p10 = os.path.join(carpeta, F10)
    wb = openpyxl.load_workbook(p10)
    wv = openpyxl.load_workbook(p10, data_only=True)['Calculadora PVP']
    ws = wb['Calculadora PVP']
    tabla = []
    for i, (nombre, mn, mx, com) in enumerate(TIPOS_10):
        f = 9 + i
        tabla.append({
            'ref': f'{F10}:Calculadora PVP:{f}', 'tipo': nombre,
            'fc_min': mn, 'fc_max': mx, 'comision': com,
            'multiplicador': wv.cell(row=f, column=5).value,
            'pvp_sin_iva': wv.cell(row=f, column=7).value,
            'pvp_con_iva': wv.cell(row=f, column=8).value,
            'ingreso_neto': wv.cell(row=f, column=9).value,
            'margen_neto': wv.cell(row=f, column=10).value,
        })
    xl = _pycel(p10)
    f_del = 9 + len(TIPOS_10) - 1
    antes = _ev(xl, "'Calculadora PVP'!G9")
    xl.set_value("'Calculadora PVP'!C4", COSTE_10 * 2)
    despues = _ev(xl, "'Calculadora PVP'!G9")
    if not (isinstance(antes, (int, float)) and isinstance(despues, (int, float))
            and abs(despues - antes * 2) < 1e-6):
        fallos.append(f'10: doblar el coste no dobla el PVP ({antes} → {despues})')
    fuera['calculadora_10'] = {
        'coste_por_racion': ws['C4'].value, 'iva': ws['C5'].value,
        'columnas': [ws.cell(row=8, column=c).value for c in range(2, 11)],
        'tabla': tabla,
        'formula_pvp': ws.cell(row=9, column=7).value,
        'formula_multiplicador': ws.cell(row=9, column=5).value,
        'sensibilidad_coste': {'coste': COSTE_10, 'pvp_fine_dining': antes,
                               'coste_x2': COSTE_10 * 2,
                               'pvp_fine_dining_x2': despues},
        'delivery': tabla[-1],
        'nota': ('El PVP de Delivery sale alto A PROPÓSITO: la comisión se '
                 'descuenta antes, y el food cost objetivo 28-32 % (SPEC §7.1) '
                 'se mide sobre el ingreso NETO — un 30 % sobre neto ≈ 21 % '
                 'sobre el PVP bruto. Con la comisión a 0 % el PVP baja a '
                 'coste × multiplicador.'),
    }

    # ---- 11 -------------------------------------------------------------
    p11 = os.path.join(carpeta, F11)
    wb = openpyxl.load_workbook(p11)
    ws = wb['Dashboard']
    wv = openpyxl.load_workbook(p11, data_only=True)['Dashboard']
    cf11 = {str(r.sqref): [type(x).__name__ for x in r.rules]
            for r in ws.conditional_formatting}
    meses = []
    for i, mes in enumerate(MESES):
        f = 7 + i
        meses.append({
            'ref': f'{F11}:Dashboard:{f}', 'mes': mes,
            'stock_inicial': ws.cell(row=f, column=3).value,
            'compras': ws.cell(row=f, column=4).value,
            'stock_final': ws.cell(row=f, column=5).value,
            'consumo': wv.cell(row=f, column=6).value,
            'ventas': ws.cell(row=f, column=7).value,
            'food_cost': wv.cell(row=f, column=8).value,
            'objetivo': wv.cell(row=f, column=9).value,
            'estado': wv.cell(row=f, column=11).value,
        })
    ene = meses[0]
    if not (ene['consumo'] == 4800 and abs((ene['food_cost'] or 0) - 0.30) < 1e-9
            and ene['estado'] == 'OK' and abs((ene['objetivo'] or 0) - 0.30) < 1e-9):
        fallos.append('11: Enero (1.000 + 5.000 − 1.200 sobre 16.000) no da '
                      f'consumo 4.800, food cost 30 % y OK → {ene}')
    if meses[1]['estado'] != 'ALERTA':
        fallos.append('11: Febrero al 36,7 % no da ALERTA → '
                      f'{meses[1]["estado"]!r}')
    # TEC-13 + TEC-R2-02: los meses sin ventas devuelven #N/A a propósito —
    # es el ÚNICO valor que el gráfico salta de verdad ("" se dibuja como 0).
    for i in range(3, 12):
        v = meses[i]['objetivo']
        if isinstance(v, (int, float)):
            fallos.append(f'11: {MESES[i]} sin ventas pinta serie objetivo '
                          f'({v!r}) — TEC-13')
            break
        if v not in (None, '#N/A'):
            fallos.append(f'11: {MESES[i]} sin ventas debería devolver #N/A y '
                          f'devuelve {v!r} — TEC-R2-02')
            break
    graf11 = _grafico(p11, 'Dashboard')
    if graf11['graficos'] != 1:
        fallos.append(f'11: {graf11["graficos"]} gráficos en Dashboard')
    for s in graf11['series']:
        if not s['categorias_strRef'] or s['categorias_numRef']:
            fallos.append('11: el gráfico sigue con las categorías como '
                          'numRef (COM-24)')
    if graf11.get('x_axis_axPos') != 'b':
        fallos.append('11: el eje de categorías no está abajo (COM-24)')
    xl = _pycel(p11)
    # pycel exige que la dirección esté en el cell map ANTES de set_value: se
    # evalúan primero las fórmulas que dependen de ella.
    for ref in ("'Dashboard'!F10", "'Dashboard'!H10", "'Dashboard'!I10",
                "'Dashboard'!K10"):
        _ev(xl, ref)
    xl.set_value("'Dashboard'!G10", 20000)
    xl.set_value("'Dashboard'!C10", 1300)
    xl.set_value("'Dashboard'!D10", 7000)
    xl.set_value("'Dashboard'!E10", 1100)
    viva11 = {'mes': 'Abril', 'se_escribe': {'C10': 1300, 'D10': 7000,
                                             'E10': 1100, 'G10': 20000},
              'consumo': _ev(xl, "'Dashboard'!F10"),
              'food_cost': _ev(xl, "'Dashboard'!H10"),
              'objetivo': _ev(xl, "'Dashboard'!I10"),
              'estado': _ev(xl, "'Dashboard'!K10")}
    if viva11['estado'] != 'ALERTA':
        fallos.append(f'11: Abril al 36 % debería dar ALERTA → {viva11}')
    fuera['dashboard_11'] = {
        'objetivo': ws['D4'].value,
        'ancho_columna_B': ws.column_dimensions['B'].width,
        'rotulo_B4': ws['B4'].value,
        'columnas': [ws.cell(row=6, column=c).value for c in range(2, 12)],
        'formula_food_cost': ws['H7'].value,
        'formula_objetivo': ws['I7'].value,
        'formula_estado': ws['K7'].value,
        'formato_condicional': cf11,
        'meses': meses,
        'total_anual': {'compras': wv['D19'].value, 'consumo': wv['F19'].value,
                        'ventas': wv['G19'].value,
                        'food_cost': wv['H19'].value},
        'grafico': graf11,
        'prueba_viva_mes_vacio': viva11,
    }

    # ---- BONUS ----------------------------------------------------------
    pbo = os.path.join(carpeta, FBO)
    wb = openpyxl.load_workbook(pbo)
    wv = openpyxl.load_workbook(pbo, data_only=True)
    inv, invv = wb['Inventario'], wv['Inventario']
    ven, venv = wb['Ventas del periodo'], wv['Ventas del periodo']
    fila_tot_ventas = 5 + FILAS_VENTAS
    productos = []
    for i, (prod, ud, ini, com, fin, precio) in enumerate(INVENTARIO):
        f = 5 + i
        productos.append({
            'ref': f'{FBO}:Inventario:{f}', 'producto': prod, 'ud': ud,
            'stock_inicial': ini, 'compras': com, 'stock_final': fin,
            'consumo_real': invv.cell(row=f, column=6).value,
            'consumo_teorico': invv.cell(row=f, column=7).value,
            'diferencia_ud': invv.cell(row=f, column=8).value,
            'precio_unitario': precio,
            'valor_diferencia_eur': invv.cell(row=f, column=10).value,
        })
    sin_ventas = [p for p in productos if p['producto'] == 'Harina'][0]
    if sin_ventas['consumo_teorico'] is not None or \
            sin_ventas['valor_diferencia_eur'] is not None:
        fallos.append('BONUS: la guarda de vacío no funciona — Harina, sin '
                      f'ventas ni stock, muestra {sin_ventas}')
    solomillo = productos[0]
    if solomillo['consumo_teorico'] != 26.4:
        fallos.append('BONUS: 120 raciones × 0,22 kg debería dar 26,40 kg de '
                      f'consumo teórico y da {solomillo["consumo_teorico"]!r}')
    checklist = wb['Checklist Mermas']
    dv_motivo = [str(dv.sqref) for dv in checklist.data_validations.dataValidation]
    if f'D4:D{3 + FILAS_REGISTRO}' not in dv_motivo:
        fallos.append(f'BONUS: el desplegable de motivo no cubre las '
                      f'{FILAS_REGISTRO} filas → {dv_motivo}')
    fuera['inventario_bonus'] = {
        'hojas': wb.sheetnames,
        'columnas_inventario': [inv.cell(row=4, column=c).value
                                for c in range(1, 11)],
        'formula_consumo_real': inv['F5'].value,
        'formula_consumo_teorico': inv['G5'].value,
        'formula_valor_diferencia': inv['J5'].value,
        'formula_sumproduct': ven.cell(row=fila_tot_ventas, column=3).value,
        'productos': productos,
        'total_desvio_eur': invv.cell(row=5 + len(INVENTARIO), column=10).value,
        'ventas_del_periodo': [{
            'plato': ven.cell(row=5 + i, column=1).value,
            'raciones': ven.cell(row=5 + i, column=2).value,
            'receta': {INVENTARIO[j][0]: ven.cell(row=5 + i, column=3 + j).value
                       for j in range(len(INVENTARIO))
                       if ven.cell(row=5 + i, column=3 + j).value is not None},
        } for i in range(len(VENTAS))],
        'consumo_teorico_totales': {
            INVENTARIO[j][0]: venv.cell(row=fila_tot_ventas, column=3 + j).value
            for j in range(len(INVENTARIO))},
        'registro_incidencias': {
            'titulo': checklist['A1'].value,
            'filas': FILAS_REGISTRO,
            'primera': 4, 'ultima': 3 + FILAS_REGISTRO,
            'total': checklist.cell(row=4 + FILAS_REGISTRO, column=6).value,
            'dv_motivo': dv_motivo,
            'formato_fecha': checklist['A4'].number_format,
        },
    }

    fuera['fallos'] = fallos
    return fuera
