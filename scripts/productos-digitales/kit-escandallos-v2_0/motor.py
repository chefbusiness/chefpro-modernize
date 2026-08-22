#!/usr/bin/env python3
"""
motor.py — Motor común de los escandallos del Kit de Escandallos Pro v2.0.

Implementa el §1 de `kit-escandallos-v2-SPEC.md` sobre las 8 plantillas de
escandallo (01…08). NO toca ficheros: recibe un `Workbook` ya cargado y lo
modifica en memoria; quien guarda es `main.py`.

Qué hace (§1.1-§1.8):
  1.1  Columna «Factor» entre «Ud. Uso» y «Merma» + hoja auxiliar `Conversiones`.
       Coste = Cant. Bruta × Precio/Ud ÷ Factor. Las 17 filas precargadas con
       Ud. Compra ≠ Ud. Uso se corrigen (tabla PARCHES).
  1.2  «Tipo de IVA (%)» en celda verde; el PVP con IVA pasa de *1.10 a *(1+IVA).
  1.3  Hoja auxiliar `Mermas` (categoría → típica/mín/máx) y VLOOKUP de merma en
       las filas VACÍAS. Las precargadas conservan su valor escrito.
  1.4  Bloque «Nº de raciones» / «COSTE POR RACIÓN» / «PVP actual en carta» /
       «FOOD COST REAL (%)» con formato condicional rojo sobre el objetivo.
  1.5  5 filas libres más por escandallo, dentro del SUM.
  1.6  Protección de hoja SIN contraseña con las celdas verdes desbloqueadas.
  1.7  Ejemplos realistas (microgreens, tostada de aguacate, colorante…).
  1.8  Instrucciones reescritas para describir EXACTAMENTE la hoja.

IDEMPOTENTE: el único paso destructivo (insertar la columna Factor + las 5 filas)
va detrás de un centinela — la cabecera G4 == 'Factor'. Todo lo demás es
escritura ABSOLUTA (siempre el mismo valor en la misma celda), así que la 2.ª
pasada deja el fichero byte-equivalente salvo timestamps del zip.

Las utilidades `_traducir_formula`, `_rangos_dv`, `_restaurar_dv`,
`_desplazar_rango`, `insertar_columna`, `insertar_fila`, `print_setup` y
`linea_instrucciones` son las de `kit-pasteleria-v2_0-postprocess.py`
(referencia probada en producción el 2026-08-22); se copian aquí para no
importar un script con `main()` y rutas propias.
"""
import copy
import re

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Paleta y formatos del kit (se mantienen los de v1.1)
# ==========================================================================
VERDE = 'E8F5E9'        # celda editable
GRIS = 'F2F2F2'         # columna calculada auxiliar (Factor)
CREMA = 'FFF8DC'        # subtotal
ORO = 'FFD700'          # total / PVP
ROJO_BG = 'FFC7CE'
ROJO_FG = '9C0006'
CAB = '2D2D2D'

FMT_EUR = '#,##0.00 €'
FMT_PCT = '0%'
FMT_PCT1 = '0.0%'
FMT_CANT = '0.000'
FMT_FACTOR = '0.###'
FMT_ENT = '0'

VERSION_LINE = ('Versión 2.0 · agosto 2026 · aichef.pro/kit-escandallos · '
                'info@aichef.pro')
RX_VERSION = re.compile(r'^Versi[óo]n \d+\.\d+ · .*kit-escandallos')

CAB_INGREDIENTE = 'Ingrediente'
CAB_TOTAL = 'COSTE TOTAL INGREDIENTES'
CAB_FACTOR = 'Factor'
ETIQ_RACIONES = 'Nº de raciones'
ETIQ_IVA = 'Tipo de IVA (%)'
ETIQ_FC = 'Food Cost objetivo (%)'
ETIQ_PVP_ACTUAL = 'PVP actual en carta (sin IVA)'
ETIQ_FC_REAL = 'FOOD COST REAL (%)'
ETIQ_COSTE_RACION = 'COSTE POR RACIÓN'
MARCA_CF = 'kitesc-v2-fcreal'

# Rótulos que el §3 (grupo B) renombra por contexto: un gin tonic no tiene
# «COSTE TOTAL DEL PLATO» y una tanda de croissants no tiene «raciones».
# El motor TIENE que seguir reconociéndolos como el mismo concepto, porque
# `bloque()` los busca por etiqueta y `_reconstruir_cola` saca de ellos DOS
# cosas: la plantilla de estilo de cada banda y el valor que ya había escrito
# el cliente. Sin este mapa la 2.ª pasada busca «PVP SUGERIDO», no lo
# encuentra, cae al desplazamiento fijo r_tot+5 —que en v2 es la fila VACÍA—
# y repinta el PVP con el estilo de una fila en blanco: exactamente el fallo
# de idempotencia (28 diferencias) que ya se cazó una vez en este motor.
ALIAS_ROTULOS = {
    'COSTE TOTAL DEL CÓCTEL': 'COSTE TOTAL DEL PLATO',       # 04
    'COSTE TOTAL DE LA TANDA': 'COSTE TOTAL DEL PLATO',      # 05
    'COSTE TOTAL POR PERSONA': 'COSTE TOTAL DEL PLATO',      # 06
    'Merma y hielo (%)': 'Coste elaboración (%)',            # 04
    'Rendimiento (uds por tanda)': ETIQ_RACIONES,            # 05
    'COSTE POR UNIDAD': ETIQ_COSTE_RACION,                   # 05
    'PVP POR UNIDAD (sin IVA)': 'PVP SUGERIDO (sin IVA)',    # 05
    'PVP POR UNIDAD (con IVA)': 'PVP CON IVA',               # 05
}


def _canon(etiqueta):
    """Etiqueta canónica del bloque de resultado.

    Normaliza las dos formas en que una etiqueta puede venir «decorada»:
    el sufijo « — …» que le añade grupo_a al food cost objetivo y los
    renombrados por contexto de grupo_b (ALIAS_ROTULOS).
    """
    if not isinstance(etiqueta, str):
        return etiqueta
    base = etiqueta.split(' — ')[0].strip()
    return ALIAS_ROTULOS.get(base, base)

FICHEROS = [
    '01-escandallo-estandar.xlsx',
    '02-menu-degustacion.xlsx',
    '03-menu-del-dia.xlsx',
    '04-cocktails-bebidas.xlsx',
    '05-pasteleria.xlsx',
    '06-catering.xlsx',
    '07-cafeteria-brunch.xlsx',
    '08-food-truck.xlsx',
]

FILAS_LIBRES_NUEVAS = 5

# Registro de TODA fórmula que escribe el motor: main.py lo usa para verificar
# una por una que quedaron con valor cacheado (o que son "" por diseño).
REGISTRO = []


def _reg(ws, coord, formula):
    REGISTRO.append((ws.title, coord, formula))

# ==========================================================================
# §1.1 — Tabla de conversiones (Ud. Compra → Ud. Uso)
# ==========================================================================
# Sólo parejas construibles con el desplegable real de las plantillas:
# kg,g,L,ml,cl,ud,docena,manojo,sobre,lata,botella.  Lo que no esté aquí vale 1
# (IFERROR), que es el comportamiento seguro: nunca multiplica ni divide de más.
CONVERSIONES = [
    ('kg→kg', 1, 'misma unidad'),
    ('kg→g', 1000, '1 kg = 1.000 g'),
    ('g→g', 1, 'misma unidad'),
    ('g→kg', 0.001, '1 g = 0,001 kg'),
    ('L→L', 1, 'misma unidad'),
    ('L→ml', 1000, '1 L = 1.000 ml'),
    ('L→cl', 100, '1 L = 100 cl'),
    ('ml→ml', 1, 'misma unidad'),
    ('ml→L', 0.001, '1 ml = 0,001 L'),
    ('ml→cl', 0.1, '1 ml = 0,1 cl'),
    ('cl→cl', 1, 'misma unidad'),
    ('cl→ml', 10, '1 cl = 10 ml'),
    ('cl→L', 0.01, '1 cl = 0,01 L'),
    ('ud→ud', 1, 'misma unidad'),
    ('docena→ud', 12, '1 docena = 12 ud'),
    ('docena→docena', 1, 'misma unidad'),
    ('ud→docena', 0.0833333333333333, '1 ud = 1/12 docena'),
    ('manojo→manojo', 1, 'misma unidad'),
    ('manojo→ud', 1, 'el manojo se compra y se usa como unidad'),
    ('sobre→sobre', 1, 'misma unidad'),
    ('sobre→ud', 1, 'el sobre se compra y se usa como unidad'),
    ('lata→lata', 1, 'misma unidad'),
    ('lata→ud', 1, 'la lata se compra y se usa como unidad'),
    ('botella→botella', 1, 'misma unidad'),
    ('botella→ud', 1, 'la botella se compra y se usa como unidad'),
    # Parejas de FORMATO DE COMPRA (DOM-01/TEC-R2-01). Sin ellas el VLOOKUP no
    # encontraba «botella→cl» —como se compra de verdad en una barra— y el
    # factor caía a 1 en silencio: el coste salía multiplicado por 70.
    # El tamaño del formato es un SUPUESTO: la columna «Factor» es editable
    # (esta hoja va sin protección) y la nota de arriba lo dice.
    ('botella→cl', 70, 'botella de 70 cl (destilados). Vino: cambia el 70 por 75'),
    ('botella→ml', 700, 'botella de 70 cl = 700 ml'),
    ('botella→L', 0.7, 'botella de 70 cl = 0,7 L'),
    ('lata→cl', 33, 'lata de 33 cl'),
    ('lata→ml', 330, 'lata de 33 cl = 330 ml'),
    ('lata→L', 0.33, 'lata de 33 cl = 0,33 L'),
    ('manojo→g', 30, 'manojo de hierbas ≈ 30 g'),
    ('sobre→g', 10, 'sobre ≈ 10 g'),
]

# ==========================================================================
# §1.3 — Mermas por categoría (típica / mín / máx, en tanto por uno)
# ==========================================================================
# Las típicas de las 16 categorías originales son las que ya publicaba
# 01!Instrucciones!B28-B30; el resto son horquillas de despiece de manual de
# cocina.  Las 5 categorías nuevas (marcadas) las exige DOM-23/TEC-26: sin
# ellas hay filas que no tienen dónde clasificarse y acaban en una familia
# ajena, que es justo lo que rompería este VLOOKUP.
MERMAS = [
    ('Carne roja', 0.20, 0.15, 0.25),
    ('Aves', 0.25, 0.20, 0.30),
    ('Pescado', 0.35, 0.30, 0.45),
    ('Marisco', 0.45, 0.35, 0.55),
    ('Verdura hoja', 0.25, 0.20, 0.35),
    ('Verdura raíz', 0.12, 0.08, 0.18),
    ('Verdura fruto', 0.10, 0.06, 0.15),          # nueva (DOM-23/TEC-26)
    ('Fruta', 0.15, 0.10, 0.25),
    ('Lácteos', 0.03, 0.01, 0.05),
    ('Secos/granos', 0.02, 0.01, 0.04),
    ('Congelados', 0.07, 0.03, 0.10),
    ('Pan/bollería', 0.05, 0.02, 0.10),
    ('Huevos', 0.11, 0.10, 0.13),
    ('Aceites/grasas', 0.03, 0.01, 0.05),
    ('Especias/hierbas', 0.20, 0.10, 0.30),
    ('Chocolate/cacao', 0.05, 0.02, 0.08),
    ('Bebidas/licores', 0.02, 0.01, 0.04),
    ('Setas/hongos', 0.20, 0.12, 0.30),           # nueva (DOM-23)
    ('Conservas/encurtidos', 0.08, 0.04, 0.15),   # nueva (DOM-23)
    ('Salsas/condimentos', 0.05, 0.02, 0.08),     # nueva (DOM-23)
    ('Pasta/arroz', 0.02, 0.01, 0.04),            # nueva (DOM-23)
]
CATEGORIAS = [m[0] for m in MERMAS]
UNIDADES = ['kg', 'g', 'L', 'ml', 'cl', 'ud', 'docena', 'manojo',
            'sobre', 'lata', 'botella']
# COM-M15: unidades en las que se CUENTAN piezas enteras y categorías cuya
# merma típica es de despiece/cáscara (sólo aplicable al peso).
UNIDADES_DISCRETAS = {'ud', 'docena'}
CATEGORIAS_POR_PIEZA = {'Huevos'}

# ==========================================================================
# §1.1 + §1.7 — Parches sobre las filas precargadas
# ==========================================================================
# Clave: (fichero, hoja, fila).  Valor: {columna_nueva: valor, ...} con las
# letras del layout v2 (A..F datos, H merma).  «motivo» documenta el hallazgo.
#
# Criterio de las 17 filas con Ud. Compra ≠ Ud. Uso (DOM-05/TEC-01/COM-04):
#   · L→cl (destilados): el número era correcto en LITROS y la etiqueta mentía.
#     Se pasa el número a cl (0,05 L → 5 cl) y el euro NO cambia: era el importe
#     bueno mal rotulado.
#   · docena→ud (huevos): el número estaba en docenas.  Se pasa a unidades
#     reales de huevo; el euro no cambia salvo donde además había un error de
#     receta (07 tostada: 2 huevos en un plato que se llama «con Huevo Poché»).
#   · manojo→ud (hierbas de guarnición): «0,5 ud de romero» no existía como
#     concepto — el desplegable no tiene «rama».  Se homogeneiza a la unidad de
#     compra (manojo) con la fracción real que se usa (1 rama ≈ 1/12 manojo) y
#     el euro BAJA, que es la corrección: 0,625 € de romero en un plato era el
#     error.  Factor 1, sin conversión inventada.
PARCHES = {
    # ---- 01 -----------------------------------------------------------
    ('01-escandallo-estandar.xlsx', 'Escandallo', 10): (
        {'B': 'Salsas/condimentos'},
        'TEC-26: un fondo oscuro no es un congelado; se reclasifica.'),
    ('01-escandallo-estandar.xlsx', 'Escandallo', 14): (
        {'A': 'Microgreens (bandeja 50 g)', 'C': 'kg', 'D': 70, 'E': 0.003,
         'F': 'kg', 'H': 0},
        'DOM-21: 3,89 € de decoración (25 % del plato) porque se tomaba el '
        'precio de la bandeja como el de la ración. Bandeja de 50 g a 3,50 € = '
        '70 €/kg; 3 g por plato = 0,21 €. Y sin merma sobre una unidad entera.'),

    # ---- 02 -----------------------------------------------------------
    ('02-menu-degustacion.xlsx', '2. Entrante', 7): (
        {'B': 'Setas/hongos'},
        'DOM-23: la trufa negra estaba en «Verdura raíz».'),
    ('02-menu-degustacion.xlsx', '4. Carne', 9): (
        {'A': 'Romero fresco (1 rama ≈ 1/12 manojo)', 'E': 0.083, 'F': 'manojo'},
        'DOM-05: C=manojo / F=ud sin factor cargaba medio manojo (0,63 €) de '
        'romero en una carrillera.'),

    # ---- 03 -----------------------------------------------------------
    ('03-menu-del-dia.xlsx', 'Primer Plato', 6): (
        {'B': 'Verdura fruto'},
        'DOM-23: el tomate pera estaba en «Verdura hoja» (25 % de merma).'),
    ('03-menu-del-dia.xlsx', 'Primer Plato', 7): (
        {'B': 'Verdura fruto'},
        'TEC-26: el pepino no es una verdura de raíz.'),
    ('03-menu-del-dia.xlsx', 'Primer Plato', 9): (
        {'B': 'Conservas/encurtidos'},
        'DOM-23: las aceitunas negras estaban en «Congelados».'),
    ('03-menu-del-dia.xlsx', 'Segundo Plato', 7): (
        {'B': 'Verdura fruto'},
        'TEC-26: el pimiento rojo no es verdura de hoja.'),
    ('03-menu-del-dia.xlsx', 'Segundo Plato', 9): (
        {'A': 'Romero fresco (1 rama ≈ 1/12 manojo)', 'E': 0.083, 'F': 'manojo'},
        'DOM-05: C=manojo / F=ud sin factor.'),
    ('03-menu-del-dia.xlsx', 'Postre', 5): (
        {'A': 'Huevo campero', 'E': 0.5, 'F': 'ud'},
        'DOM-05: 0,25 «ud» que en realidad eran 0,25 DOCENAS → 1,01 € de huevo '
        'en un flan de menú del día. Media unidad por ración (4 huevos por '
        'litro de leche, 8 raciones) = 0,17 €.'),

    # ---- 04 -----------------------------------------------------------
    ('04-cocktails-bebidas.xlsx', 'Gin Tonic Premium', 5): (
        {'E': 5},
        'DOM-05: 0,05 L etiquetados «cl». Son 5 cl; el euro no cambia.'),
    ('04-cocktails-bebidas.xlsx', 'Gin Tonic Premium', 8): (
        {'B': 'Verdura fruto'},
        'TEC-26: el pepino estaba en «Verdura raíz».'),
    ('04-cocktails-bebidas.xlsx', 'Mojito Clásico', 5): (
        {'E': 6}, 'DOM-05: 0,06 L etiquetados «cl» → 6 cl de ron.'),
    ('04-cocktails-bebidas.xlsx', 'Mojito Clásico', 7): (
        {'A': 'Hierbabuena (1 rama ≈ 1/12 manojo)', 'E': 0.083, 'F': 'manojo'},
        'DOM-05: medio manojo de hierbabuena (0,75 €) en un mojito.'),
    ('04-cocktails-bebidas.xlsx', 'Margarita', 5): (
        {'E': 5}, 'DOM-05: 0,05 L etiquetados «cl» → 5 cl de tequila.'),
    ('04-cocktails-bebidas.xlsx', 'Margarita', 6): (
        {'E': 2.5}, 'DOM-05: 0,025 L etiquetados «cl» → 2,5 cl de triple seco.'),
    ('04-cocktails-bebidas.xlsx', 'Aperol Spritz', 5): (
        {'E': 6}, 'DOM-05: 0,06 L etiquetados «cl» → 6 cl de Aperol.'),
    ('04-cocktails-bebidas.xlsx', 'Aperol Spritz', 6): (
        {'E': 9}, 'DOM-05: 0,09 L etiquetados «cl» → 9 cl de prosecco.'),
    ('04-cocktails-bebidas.xlsx', 'Aperol Spritz', 7): (
        {'E': 3}, 'DOM-05: 0,03 L etiquetados «cl» → 3 cl de soda.'),

    # ---- 05 -----------------------------------------------------------
    ('05-pasteleria.xlsx', 'Tarta Chocolate', 7): (
        {'A': 'Huevos (unidades)', 'E': 6},
        'DOM-05: 0,5 «ud» que eran 0,5 docenas. Son 6 huevos; el euro no cambia.'),
    ('05-pasteleria.xlsx', 'Croissants', 9): (
        {'B': 'Secos/granos'},
        'DOM-23: la levadura fresca estaba clasificada como «Lácteos».'),
    ('05-pasteleria.xlsx', 'Croissants', 11): (
        {'E': 1},
        'DOM-05: 0,083 «ud» que eran 0,083 docenas = 1 huevo de pintar.'),
    ('05-pasteleria.xlsx', 'Macarons', 11): (
        {'A': 'Colorante rojo (bote 50 g · 2 g/tanda)', 'B': 'Secos/granos',
         'E': 0.04},
        'TEC-26: 0,1 «ud» de colorante daba 100 g de colorante para 30 macarons '
        'sobre 500 g de masa. La dosis real son 2 g = 0,04 botes de 50 g.'),

    # ---- 07 -----------------------------------------------------------
    ('07-cafeteria-brunch.xlsx', 'Tostada Aguacate', 5): (
        {'A': 'Pan de masa madre (rebanada gruesa)', 'E': 1},
        'DOM-20: 2 rebanadas (1,20 € sólo de pan) en una tostada.'),
    ('07-cafeteria-brunch.xlsx', 'Tostada Aguacate', 6): (
        {'A': 'Aguacate Hass (½ pieza)', 'E': 0.4},
        'DOM-20: un aguacate ENTERO, y encima con 28 % de merma sobre la unidad '
        'ya contada. Media pieza neta = 0,4 ud.'),
    ('07-cafeteria-brunch.xlsx', 'Tostada Aguacate', 7): (
        {'A': 'Huevo campero', 'E': 1, 'F': 'ud'},
        'DOM-20: 0,167 «ud» eran 0,167 docenas = 2 huevos en un plato que se '
        'llama «con Huevo Poché», en singular.'),
    ('07-cafeteria-brunch.xlsx', 'Tostada Aguacate', 8): (
        {'B': 'Verdura fruto'},
        'TEC-26: el tomate cherry estaba en «Verdura hoja».'),
    ('07-cafeteria-brunch.xlsx', 'Eggs Benedict', 6): (
        {'A': 'Huevo campero', 'E': 2, 'F': 'ud'},
        'DOM-05: 0,167 «ud» eran 0,167 docenas = los 2 huevos que sí lleva un '
        'eggs benedict. El euro no cambia.'),
    ('07-cafeteria-brunch.xlsx', 'Eggs Benedict', 10): (
        {'A': 'Cebollino (1 rama ≈ 1/12 manojo)', 'E': 0.083, 'F': 'manojo'},
        'DOM-05: 0,25 «ud» de un manojo → 0,47 € de cebollino de decoración.'),

    # ---- 08 -----------------------------------------------------------
    ('08-food-truck.xlsx', 'Smash Burger', 10): (
        {'B': 'Verdura fruto'},
        'TEC-26: el tomate llevaba 10 % de merma por estar en «Verdura hoja» '
        'mientras la lechuga de al lado llevaba 25 %.'),
    ('08-food-truck.xlsx', 'Smash Burger', 11): (
        {'B': 'Salsas/condimentos'},
        'DOM-23: la salsa especial estaba en «Secos/granos».'),
    ('08-food-truck.xlsx', 'Smash Burger', 12): (
        {'B': 'Conservas/encurtidos'},
        'DOM-23: el pepinillo estaba en «Congelados».'),
    ('08-food-truck.xlsx', 'Loaded Fries', 9): (
        {'A': 'Cebollino (1 rama ≈ 1/12 manojo)', 'E': 0.083, 'F': 'manojo'},
        'DOM-05: 0,125 «ud» de un manojo sin factor.'),
    ('08-food-truck.xlsx', 'Pulled Pork Sándwich', 9): (
        {'B': 'Conservas/encurtidos'},
        'DOM-23: los pepinillos estaban en «Congelados» (coherencia con el 08 '
        'Smash Burger).'),
}

# Las 17 filas del kit en que Ud. Compra ≠ Ud. Uso (censo DOM-05). Se usan para
# el informe: fichero:hoja:fila con coste antes/después.
FILAS_UNIDAD_MIXTA = [
    ('02-menu-degustacion.xlsx', '4. Carne', 9),
    ('03-menu-del-dia.xlsx', 'Segundo Plato', 9),
    ('03-menu-del-dia.xlsx', 'Postre', 5),
    ('04-cocktails-bebidas.xlsx', 'Gin Tonic Premium', 5),
    ('04-cocktails-bebidas.xlsx', 'Mojito Clásico', 5),
    ('04-cocktails-bebidas.xlsx', 'Mojito Clásico', 7),
    ('04-cocktails-bebidas.xlsx', 'Margarita', 5),
    ('04-cocktails-bebidas.xlsx', 'Margarita', 6),
    ('04-cocktails-bebidas.xlsx', 'Aperol Spritz', 5),
    ('04-cocktails-bebidas.xlsx', 'Aperol Spritz', 6),
    ('04-cocktails-bebidas.xlsx', 'Aperol Spritz', 7),
    ('05-pasteleria.xlsx', 'Tarta Chocolate', 7),
    ('05-pasteleria.xlsx', 'Croissants', 11),
    ('07-cafeteria-brunch.xlsx', 'Tostada Aguacate', 7),
    ('07-cafeteria-brunch.xlsx', 'Eggs Benedict', 6),
    ('07-cafeteria-brunch.xlsx', 'Eggs Benedict', 10),
    ('08-food-truck.xlsx', 'Loaded Fries', 9),
]

# ==========================================================================
# §1.8 — Instrucciones
# ==========================================================================
PROPIAS = {
    '01-escandallo-estandar.xlsx': [
        "▸ Ve a la pestaña 'Escandallo' para empezar.",
        '▸ El ejemplo cargado es un solomillo con Pedro Ximénez: bórralo o '
        'escribe encima.',
    ],
    '02-menu-degustacion.xlsx': [
        '▸ Cada pestaña numerada es un pase del menú: hay NUEVE '
        "('1. Aperitivo'…'9. Pase'), los cinco primeros con ejemplo y los "
        'cuatro últimos vacíos.',
        "▸ La pestaña 'Resumen' suma sólo los pases que tengan ingredientes: "
        'los que dejes vacíos no ensucian el total.',
        '▸ El Food Cost objetivo se edita UNA sola vez, en el Resumen; cada '
        'pase lo lee de allí, así que no hay dos precios para lo mismo.',
    ],
    '03-menu-del-dia.xlsx': [
        "▸ Pestañas 'Primer Plato', 'Segundo Plato' y 'Postre': un escandallo "
        'cada una.',
        "▸ Pestaña 'Resumen Menú': coste del menú completo (con pan, bebida y "
        'café) y PVP sugerido.',
        "▸ Pestaña 'Rotación Semanal': los cinco menús de lunes a viernes con "
        'su coste, su PVP y su food cost por día, más la media de la semana.',
        '▸ El Food Cost objetivo se edita UNA sola vez, en el Resumen Menú.',
    ],
    '04-cocktails-bebidas.xlsx': [
        '▸ Una pestaña por cóctel: gin tonic, mojito, margarita y Aperol '
        'spritz.',
        '▸ Las cantidades de destilados van en CL y el precio en €/L: el '
        'Factor (columna G) hace la conversión. 5 cl de una ginebra de 28 €/L '
        'son 1,40 € de producto (1,43 € en la hoja, porque esa fila lleva '
        'además un 2 % de merma de servicio), no 140 €.',
        '▸ Pestaña «Formatos de Compra»: convierte el precio de la BOTELLA de '
        'la factura (70 cl los destilados, 75 cl el vino y el espumoso) en el '
        '€/L que pide la columna «Precio/Ud». Es el error que más cuesta: una '
        'ginebra de 19,60 € la botella son 28 €/L.',
        # TEC-R2-15: antes había que copiar el €/L a mano y el escandallo se
        # quedaba al precio viejo en cuanto cambiaba la factura.
        '▸ El «Precio/Ud» de los destilados, el vino y el espumoso YA VIENE '
        'enlazado a «Formatos de Compra»: por eso esas celdas son azules y no '
        'verdes. Cambia el precio de la botella UNA vez allí y se actualizan '
        'todos los cócteles. Si quieres teclear el €/L a mano, desprotege la '
        'hoja y escribe encima (romperás el vínculo sólo en esa fila).',
        '▸ «Merma y hielo (%)» sustituye al coste de elaboración: el 5 % cubre '
        'el overpour de servir sin jigger y la dilución del hielo.',
        '▸ Food cost objetivo de referencia en barra: 20-25 %.',
    ],
    '05-pasteleria.xlsx': [
        '▸ Una pestaña por elaboración: tarta de chocolate (12 raciones), '
        'croissants (20 unidades) y macarons (30 unidades).',
        '▸ El escandallo es de la TANDA COMPLETA. «Rendimiento (uds por '
        'tanda)» reparte el coste: «COSTE POR UNIDAD» y «PVP POR UNIDAD» son '
        'los que van a la vitrina, no el total de la tanda.',
        '▸ Mermas propias de obrador ya cargadas: 12 % en el chocolate de '
        'cobertura (pérdida de temperado) y 8 % en las harinas (evaporación '
        'de horneado).',
        '▸ El colorante se escandalla por bote: 2 g de un bote de 50 g = 0,04 '
        'botes, no 0,1 «unidades».',
        '▸ Food cost objetivo de referencia en pastelería: 20-30 %.',
    ],
    '06-catering.xlsx': [
        '▸ La pestaña «Cocktail (por persona)» se escandalla POR COMENSAL. El '
        'número de comensales se escribe UNA sola vez, en «Presupuesto»!C5: si '
        'lo metes también en el escandallo, facturas el evento por partida '
        'doble.',
        '▸ «Presupuesto» es la hoja INTERNA: enseña tus costes y tu margen. El '
        'bloque de PVP de la hoja de escandallo es orientativo y sólo cubre la '
        'materia prima; el precio al cliente sale del Presupuesto.',
        '▸ El personal ya no es una constante escondida: camareros = '
        'REDONDEAR.MÁS(comensales ÷ comensales por camarero) × horas × €/hora, '
        'más el jefe de sala. Todos los parámetros son celdas verdes.',
        '▸ «Checklist Evento»: marca con ✓ en el desplegable de la última '
        'columna; el contador del final lleva la cuenta sobre el total.',
        '▸ «Presupuesto Cliente» es la ÚNICA hoja que se le enseña al cliente: '
        'concepto, comensales y precio con IVA. Sin costes ni margen.',
        '▸ Food cost objetivo de referencia en catering: 30-40 %.',
    ],
    '07-cafeteria-brunch.xlsx': [
        '▸ 4 recetas típicas de brunch y cafetería: tostada de aguacate, açaí '
        'bowl, eggs benedict y carrot cake.',
        '▸ La carrot cake se escandalla por tanda: 12 raciones en «Nº de '
        'raciones» y el coste por porción sale solo.',
        '▸ Food cost objetivo de referencia en cafetería: 25-30 %.',
    ],
    '08-food-truck.xlsx': [
        '▸ Una pestaña por producto de venta: smash burger, loaded fries y '
        'pulled pork.',
        '▸ Pestaña «Punto de Equilibrio»: escribe tus costes fijos del día y '
        'el mix de ventas, y la hoja te dice cuántas unidades tienes que '
        'vender para cubrirlos y con qué facturación.',
        '▸ El margen de contribución sale de las tres pestañas de escandallo: '
        'si cambias una receta o un precio, el punto de equilibrio se mueve '
        'solo.',
        '▸ Food cost objetivo de referencia en street food: 28-35 %.',
    ],
}

BLOQUE_COLUMNAS = [
    ('Columnas de la hoja de escandallo', [
        '▸ Ingrediente: el producto tal como lo compras.',
        '▸ Categoría: familia del ingrediente (desplegable). Es la que precarga '
        'la merma en las filas vacías.',
        '▸ Ud. Compra: la unidad en la que COMPRAS (kg, L, docena, manojo…).',
        '▸ Precio/Ud (€): lo que pagas por esa unidad de compra.',
        '▸ Cantidad: la que usas en el plato, EN LA UNIDAD DE USO.',
        '▸ Ud. Uso: la unidad en la que mides al cocinar (g, ml, cl, ud…).',
        '▸ Factor: conversión automática Ud. Compra → Ud. Uso, de la hoja '
        '«Conversiones». Si la pareja NO está en la tabla, el Factor se pone '
        'en rojo con un «?» y el Coste dice «revisa unidades»: añade la '
        'equivalencia en «Conversiones» antes de fiarte del número. Nunca da '
        'un coste plausible y falso.',
        '▸ Merma (%): desperdicio. En las filas de ejemplo está escrita a mano; '
        'en las vacías la trae la hoja «Mermas» en cuanto eliges categoría, y '
        'puedes escribir encima.',
        # COM-M15
        '▸ Ojo con la merma cuando mides en PIEZAS: la del 11 % de los huevos '
        'es la CÁSCARA, y sólo tiene sentido sobre el peso. Si compras y mides '
        'en unidades, déjala en 0 %: seis huevos son seis huevos, no 6,74. La '
        'misma regla vale para cualquier ingrediente que cuentes por unidad.',
        '▸ Cant. Bruta: lo que hay que comprar contando la merma (automático).',
        '▸ Coste (€): Cant. Bruta × Precio/Ud ÷ Factor (automático).',
    ]),
    ('Bloque de resultado', [
        '▸ Coste elaboración (%): recargo sobre la materia prima para cubrir '
        'energía, envasado y pérdidas de preparación que no se escandallan '
        'línea a línea. Viene al 10 %; ponlo a 0 % si prefieres imputar esos '
        'costes aparte.',
        '▸ COSTE TOTAL DEL PLATO: materia prima + coste de elaboración.',
        '▸ Nº de raciones: cuántas raciones salen de lo escandallado. Déjalo en '
        '1 si es un plato individual.',
        '▸ COSTE POR RACIÓN: coste total ÷ raciones. Es la base del PVP.',
        '▸ Food Cost objetivo (%): el que quieres conseguir. PVP sin IVA = '
        'coste por ración ÷ food cost objetivo.',
        '▸ Tipo de IVA (%): editable. 10 % en hostelería en España; cámbialo si '
        'te aplica otro tipo (IGIC en Canarias, o el IVA/ITBIS/IVU de tu país '
        'en Latinoamérica).',
        '▸ PVP actual en carta (sin IVA): lo que cobras HOY. Déjalo vacío si '
        'todavía no lo tienes.',
        '▸ FOOD COST REAL (%): coste por ración ÷ PVP actual. Se pone en rojo '
        'en cuanto supera tu objetivo.',
    ]),
    ('Hojas auxiliares', [
        '▸ «Conversiones»: equivalencias Ud. Compra → Ud. Uso. Va SIN '
        'protección: añade las parejas que te falten y corrige el tamaño de '
        'los formatos (la botella viene a 70 cl y la lata a 33 cl; si tu vino '
        'es de 75 cl, cambia el 70 por 75 en «botella→cl»).',
        '▸ «Mermas»: merma mínima, típica y máxima por categoría. La típica es '
        'la que se precarga en las filas vacías.',
    ]),
    ('Protección de la hoja', [
        '▸ Las hojas están protegidas SIN contraseña para que no borres una '
        'fórmula sin querer: sólo se escriben las celdas verdes.',
        '▸ Para tocar cualquier otra cosa —insertar filas, pegar la foto del '
        'plato, cambiar una fórmula—: Revisar → Desproteger hoja. No pide '
        'contraseña.',
        # TEC-R2-08: el total sigue siendo un SUM de rango cerrado. Si el
        # cliente escribe DEBAJO de la última fila libre, su ingrediente se
        # queda fuera del coste sin un solo aviso.
        '▸ Si te quedas sin filas: desprotege la hoja e inserta filas DENTRO '
        'del bloque de ingredientes (clic derecho sobre una fila del bloque → '
        'Insertar), nunca debajo de la última. Así el «COSTE TOTAL '
        'INGREDIENTES» las recoge solo; lo que escribas por debajo del bloque '
        'NO suma.',
        '▸ La foto del plato NO se puede arrastrar con la hoja protegida: '
        'primero Revisar → Desproteger hoja, y después Insertar → Imagen.',
    ]),
    # TEC-R2-09: la lista era LITERAL y se quedó en 16 de las 21 categorías
    # que ofrece el desplegable (faltaban pan, aceites, especias, chocolate y
    # bebidas, todas usadas por las filas de ejemplo). Ahora se construye
    # leyendo MERMAS, así que no puede volver a desincronizarse.
    ('Referencia de mermas estándar', '@mermas'),
]


def _lineas_mermas(por_linea=4):
    """Renglones «▸ Cat A 20 % | Cat B 25 % …» generados desde MERMAS."""
    fuera = []
    for i in range(0, len(MERMAS), por_linea):
        trozo = MERMAS[i:i + por_linea]
        fuera.append('▸ ' + ' | '.join('{} {} %'.format(nom, round(tip * 100))
                                       for nom, tip, _, _ in trozo))
    fuera.append(f'▸ Son las {len(MERMAS)} categorías del desplegable. La '
                 'tabla completa, con mínimo y máximo, está en la hoja '
                 '«Mermas».')
    return fuera

# DOM-08: el bloque genérico describe «Nº de raciones / COSTE POR RACIÓN /
# PVP sin IVA», rótulos que en 05-pastelería NO existen (grupo_b los renombra a
# «Rendimiento (uds por tanda) / COSTE POR UNIDAD / PVP POR UNIDAD»). Las
# Instrucciones prometían campos que no están, que es justo lo que la SPEC §1.8
# venía a arreglar. Aquí se sustituye la sección entera por fichero.
BLOQUE_OVERRIDE = {
    '05-pasteleria.xlsx': {
        'Bloque de resultado': [
            '▸ Coste elaboración (%): recargo sobre la materia prima para '
            'cubrir energía del horno, envasado y pérdidas de obrador que no '
            'se escandallan línea a línea. Viene al 10 %; ponlo a 0 % si '
            'prefieres imputar esos costes aparte.',
            '▸ COSTE TOTAL DE LA TANDA: lo que cuesta la receta completa, no '
            'una pieza. Todo el escandallo está en cantidades de TANDA.',
            '▸ Rendimiento (uds por tanda): cuántas piezas salen de la receta '
            'completa (12 tartas, 20 croissants, 30 macarons…). Es la celda '
            'verde que reparte el coste.',
            '▸ COSTE POR UNIDAD: coste de la tanda ÷ rendimiento. Es lo que te '
            'cuesta la pieza que pones en la vitrina.',
            '▸ Food Cost objetivo (%): el que quieres conseguir. PVP POR '
            'UNIDAD (sin IVA) = coste por unidad ÷ food cost objetivo.',
            '▸ Tipo de IVA (%): editable. 10 % en hostelería en España; '
            'cámbialo si te aplica otro tipo (IGIC en Canarias, o el '
            'IVA/ITBIS/IVU de tu país en Latinoamérica).',
            '▸ PVP actual en carta (sin IVA): lo que cobras HOY por PIEZA. '
            'Déjalo vacío si todavía no lo tienes.',
            '▸ FOOD COST REAL (%): coste por unidad ÷ PVP actual. Se pone en '
            'rojo en cuanto supera tu objetivo.',
        ],
    },
}

# COM-B16: las Instrucciones son texto GENÉRICO del motor y hablaban de
# «plato» tres veces en el manual de un libro de coctelería y en el de
# pastelería. Los rótulos de las hojas sí se renombran (grupo_b), las
# Instrucciones no. Se sustituyen por pares explícitos porque el sustantivo
# cambia de género («el cóctel» / «la elaboración») y una sustitución de la
# palabra suelta dejaría concordancias rotas. El rótulo en mayúsculas se
# sustituye por el que de verdad lleva la hoja (ronda 2b, COM-B16).
SUSTANTIVO = {
    '04-cocktails-bebidas.xlsx': (
        ('un plato individual', 'un cóctel individual'),
        ('en el plato', 'en el cóctel'),
        ('del plato', 'del cóctel'),
        ('COSTE TOTAL DEL PLATO', 'COSTE TOTAL DEL CÓCTEL'),   # rótulo real de las 4 pestañas (ronda 2b)
    ),
    '05-pasteleria.xlsx': (
        ('un plato individual', 'una elaboración individual'),
        ('en el plato', 'en la elaboración'),
        ('del plato', 'de la elaboración'),
        ('COSTE TOTAL DEL PLATO', 'COSTE TOTAL DE LA TANDA'),  # rótulo real de las 3 pestañas
    ),
}


def _sustantivo(fname, texto):
    for viejo, nuevo in SUSTANTIVO.get(fname, ()):
        texto = texto.replace(viejo, nuevo)
    return texto


PIE = '— Kit de Escandallos Pro · AI Chef Pro · aichef.pro'
COPY = '© 2026 AI Chef Pro · Todos los derechos reservados'
TITULO_COMO_USAR = 'Cómo usar esta plantilla'
CAPTION_FOTO = 'Insertar → Imagen (desprotege la hoja antes)'
# DOM-14: el rótulo «Arrastra aquí la foto» viene de la v1.1 y con la hoja
# protegida (v2.0) arrastrar no funciona. Se funden los DOS rótulos en uno
# solo, el que sí funciona, y se borra el segundo (CAPTION_FOTO).
RX_ARRASTRE = re.compile(r'^Arrastra aquí la foto\s*\n?\s*(.*)$', re.S)


def _texto_foto(valor):
    m = RX_ARRASTRE.match(valor)
    cola = (m.group(1) or '').strip() if m else ''
    sujeto = cola.rstrip(':').strip() if cola else 'del plato terminado'
    return ('Foto ' + sujeto + ':\nRevisar → Desproteger hoja,\n'
            'luego Insertar → Imagen')


# ==========================================================================
# Utilidades genéricas (de kit-pasteleria-v2_0-postprocess.py)
# ==========================================================================
RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _traducir_formula(valor, idx, eje):
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        ci = column_index_from_string(col)
        fi = int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        return f'{d1}{col}{d2}{fila}'

    return RX_REF.sub(_sub, valor)


def _rangos_dv(ws):
    return [({k: getattr(dv, k, None) for k in CAMPOS_DV},
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**{k: v for k, v in attrs.items() if v is not None})
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje) if idx else r)


def _desplazar_rango(ref, idx, eje):
    partes = ref.split(':')
    fuera = []
    for p in partes:
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        fuera.append(f'{d1}{col}{d2}{fila}')
    return ':'.join(fuera)


def insertar_columna(ws, idx):
    """Inserta una columna en idx manteniendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos de columna."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos = {k: v.width for k, v in ws.column_dimensions.items() if v.width}

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra, ancho in sorted(anchos.items(),
                               key=lambda kv: -column_index_from_string(kv[0])):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = ancho


def insertar_fila(ws, idx):
    """Equivalente por filas de insertar_columna."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = {k: v.height for k, v in ws.row_dimensions.items() if v.height}

    for fila in range(max_r, idx - 1, -1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila + 1, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'fila'))
    _restaurar_dv(ws, dvs, idx, 'fila')

    for fila, alto in sorted(alturas.items(), reverse=True):
        if fila >= idx:
            ws.row_dimensions[fila + 1].height = alto


def print_setup(ws, header_row=None, landscape=True):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = f'{header_row}:{header_row}'
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def linea_instrucciones(ws, texto, rx=None):
    """Escribe `texto` en Instrucciones: sustituye la línea que case con `rx`
    (línea de versión) o la añade al final si no existe. Nunca duplica."""
    col = 2 if ws.cell(row=2, column=2).value else 1
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)


# ==========================================================================
# Escaneo de una hoja de escandallo
# ==========================================================================
class Layout:
    """Coordenadas vivas de una hoja de escandallo (layout v2)."""

    def __init__(self, ws):
        self.ws = ws
        self.hdr = None
        for r in range(1, 10):
            if ws.cell(row=r, column=1).value == CAB_INGREDIENTE:
                self.hdr = r
                break
        self.r_tot = None
        if self.hdr:
            for r in range(self.hdr + 1, ws.max_row + 2):
                if ws.cell(row=r, column=1).value == CAB_TOTAL:
                    self.r_tot = r
                    break
        self.d0 = self.hdr + 1 if self.hdr else None
        self.d1 = self.r_tot - 2 if self.r_tot else None

    @property
    def valida(self):
        return self.hdr is not None and self.r_tot is not None and self.d1 >= self.d0

    @property
    def v2(self):
        return self.ws.cell(row=self.hdr, column=7).value == CAB_FACTOR


def hojas_escandallo(wb):
    """Hojas del libro que llevan la rejilla de escandallo, en orden."""
    fuera = []
    for ws in wb.worksheets:
        lay = Layout(ws)
        if lay.valida:
            fuera.append((ws, lay))
    return fuera


# ==========================================================================
# Hojas auxiliares
# ==========================================================================
def _hoja(wb, nombre, indice=None):
    if nombre in wb.sheetnames:
        return wb[nombre]
    ws = wb.create_sheet(nombre) if indice is None else wb.create_sheet(nombre, indice)
    return ws


def _cabecera(ws, fila, textos, anchos):
    negrita = Font(bold=True, color='FFFFFF')
    fondo = PatternFill('solid', fgColor=CAB)
    for i, t in enumerate(textos, start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = negrita
        c.fill = fondo
        c.alignment = Alignment(horizontal='center', vertical='center')
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a


def escribir_conversiones(wb, informe):
    nuevo = 'Conversiones' not in wb.sheetnames
    ws = _hoja(wb, 'Conversiones')
    ws.cell(row=1, column=1, value='Tabla de conversión Ud. Compra → Ud. Uso').font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value='La usa la columna «Factor» de cada escandallo. Puedes añadir '
                  'parejas al final: la clave se escribe «compra→uso».')
    _cabecera(ws, 4, ['Conversión (compra→uso)', 'Factor', 'Qué significa'],
              [26, 12, 46])
    fila = 5
    for clave, factor, nota in CONVERSIONES:
        ws.cell(row=fila, column=1, value=clave)
        c = ws.cell(row=fila, column=2, value=factor)
        c.number_format = FMT_FACTOR
        c.alignment = Alignment(horizontal='center')
        ws.cell(row=fila, column=3, value=nota)
        fila += 1
    # limpia restos de una pasada con más filas (nunca deja '' )
    for r in range(fila, ws.max_row + 1):
        for col in range(1, 4):
            ws.cell(row=r, column=col).value = None
    if nuevo:
        informe.append('hoja «Conversiones» creada '
                       f'({len(CONVERSIONES)} parejas)')
    return ws, 5, fila - 1


def escribir_mermas(wb, informe):
    nuevo = 'Mermas' not in wb.sheetnames
    ws = _hoja(wb, 'Mermas')
    ws.cell(row=1, column=1, value='Mermas de referencia por categoría').font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value='La «típica» es la que se precarga en las filas vacías del '
                  'escandallo al elegir categoría. Ajústala a tu proveedor.')
    _cabecera(ws, 4, ['Categoría', 'Merma típica', 'Mínima', 'Máxima'],
              [24, 14, 12, 12])
    fila = 5
    for cat, tip, mn, mx in MERMAS:
        ws.cell(row=fila, column=1, value=cat)
        for col, v in ((2, tip), (3, mn), (4, mx)):
            c = ws.cell(row=fila, column=col, value=v)
            c.number_format = FMT_PCT
            c.alignment = Alignment(horizontal='center')
        fila += 1
    for r in range(fila, ws.max_row + 1):
        for col in range(1, 5):
            ws.cell(row=r, column=col).value = None
    if nuevo:
        informe.append(f'hoja «Mermas» creada ({len(MERMAS)} categorías)')
    return ws, 5, fila - 1


# ==========================================================================
# Validaciones de datos
# ==========================================================================
def _dv_lista(formula, prompt):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True,
                        showDropDown=False)
    dv.showErrorMessage = True
    dv.errorTitle = 'Valor no válido'
    dv.error = prompt
    return dv


def rehacer_dv(ws, lay, cat_rango):
    """Rehace las 3 validaciones de la rejilla sobre TODO el bloque de datos.

    `cat_rango` tiene que ser de UNA sola columna: una lista de validación con
    un rango de tres columnas (Mermas!$A:$C) le ofrece al usuario las celdas de
    porcentaje además de las categorías.
    """
    ws.data_validations.dataValidation = []
    inline_cat = '"' + ','.join(CATEGORIAS) + '"'
    f_cat = inline_cat if len(inline_cat) <= 255 else cat_rango
    dv_cat = _dv_lista(f_cat, 'Elige una categoría de la lista (hoja «Mermas»).')
    dv_ud = _dv_lista('"' + ','.join(UNIDADES) + '"',
                      'Elige una unidad de la lista.')
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_ud)
    dv_cat.add(f'B{lay.d0}:B{lay.d1}')
    dv_ud.add(f'C{lay.d0}:C{lay.d1}')
    dv_ud.add(f'F{lay.d0}:F{lay.d1}')


# ==========================================================================
# Motor: una hoja de escandallo
# ==========================================================================
def _es_precargada(ws, fila):
    return isinstance(ws.cell(row=fila, column=1).value, str) and \
        ws.cell(row=fila, column=1).value.strip() != ''


def _pintar(cel, color=None, fmt=None, align=None, bold=None, locked=None):
    if color is None:
        cel.fill = PatternFill()
    else:
        cel.fill = PatternFill('solid', fgColor=color)
    if fmt:
        cel.number_format = fmt
    if align:
        cel.alignment = Alignment(horizontal=align, vertical='center')
    if bold is not None:
        f = copy.copy(cel.font)
        f.bold = bold
        cel.font = f
    if locked is not None:
        cel.protection = Protection(locked=locked)


def aplicar_hoja(ws, fname, informe, mermas_rango, conv_rango, cat_rango):
    """Aplica §1.1-§1.7 a una hoja de escandallo. Idempotente."""
    lay = Layout(ws)
    if not lay.valida:
        return None

    # ---- paso destructivo, una sola vez (centinela: cabecera «Factor») ----
    if not lay.v2:
        insertar_columna(ws, 7)                 # G = Factor
        for _ in range(FILAS_LIBRES_NUEVAS):    # 5 filas libres más (§1.5)
            insertar_fila(ws, lay.d1)
        informe.append(f'{ws.title}: columna Factor + {FILAS_LIBRES_NUEVAS} '
                       'filas libres')
        lay = Layout(ws)

    hdr, d0, d1, r_tot = lay.hdr, lay.d0, lay.d1, lay.r_tot

    # ---- cabeceras -------------------------------------------------------
    modelo = ws.cell(row=hdr, column=6)
    cf = ws.cell(row=hdr, column=7, value=CAB_FACTOR)
    cf._style = copy.copy(modelo._style)
    ws.cell(row=hdr, column=8).value = 'Merma (%)'
    ws.cell(row=hdr, column=9).value = 'Cant. Bruta'
    ws.cell(row=hdr, column=10).value = 'Coste (€)'
    ws.column_dimensions['G'].width = 8

    # ---- parches de las filas precargadas (§1.1 + §1.7) ------------------
    for fila in range(d0, d1 + 1):
        parche = PARCHES.get((fname, ws.title, fila))
        if not parche:
            continue
        campos, motivo = parche
        for letra, valor in campos.items():
            ws[f'{letra}{fila}'] = valor
        informe.append(f'{ws.title}!{fila}: {motivo}')

    # ---- COM-M15: merma de cáscara sobre PIEZAS ---------------------------
    # Al pasar las cantidades a la unidad de uso (§1.1) las filas de huevo
    # quedaron en «ud» arrastrando el 11 % de merma, que es de CÁSCARA y sólo
    # tiene sentido sobre el peso: la plantilla pedía comprar 6,74 huevos para
    # una tarta de 6, e inflaba esa línea un 12 %. Regla general, no una lista
    # de celdas: cualquier fila contada por unidades no lleva merma de despiece.
    for fila in range(d0, d1 + 1):
        if not _es_precargada(ws, fila):
            continue
        cat = ws.cell(row=fila, column=2).value
        uso = ws.cell(row=fila, column=6).value
        h = ws.cell(row=fila, column=8)
        if (cat in CATEGORIAS_POR_PIEZA and isinstance(uso, str)
                and uso.strip() in UNIDADES_DISCRETAS
                and isinstance(h.value, (int, float)) and h.value):
            informe.append(f'{ws.title}!{fila}: merma {h.value:.0%} → 0 % '
                           f'(«{cat}» medido en «{uso}»: la merma de cáscara '
                           'va sobre el peso, no sobre las piezas) [COM-M15]')
            h.value = 0

    # ---- rejilla: fórmulas y estilos -------------------------------------
    for fila in range(d0, d1 + 1):
        g = ws.cell(row=fila, column=7)
        # DOM-01/TEC-R2-01: el fallback ya NO es 1. Una pareja que no está en
        # «Conversiones» (botella→cl, ud→g, kg→ud…) devolvía factor 1 sin un
        # solo aviso y el coste salía multiplicado hasta ×1.000. Ahora marca
        # «?» en rojo y el coste dice «revisa unidades». La fila vacía (sin
        # Ud. Compra ni Ud. Uso) se queda en blanco, como antes.
        g.value = (f'=IF(C{fila}&F{fila}="","",'
                   f'IFERROR(VLOOKUP(C{fila}&"→"&F{fila},{conv_rango},2,FALSE),'
                   f'"?"))')
        _reg(ws, f'G{fila}', g.value)
        _pintar(g, GRIS, FMT_FACTOR, 'center', bold=False, locked=True)
        g.font = Font(size=9, color='595959')

        h = ws.cell(row=fila, column=8)
        if _es_precargada(ws, fila):
            if isinstance(h.value, str) and h.value.startswith('='):
                h.value = 0            # no debería pasar; se deja explícito
        else:
            h.value = f'=IFERROR(VLOOKUP(B{fila},{mermas_rango},2,FALSE),"")'
            _reg(ws, f'H{fila}', h.value)
        h.number_format = FMT_PCT

        i = ws.cell(row=fila, column=9)
        i.value = (f'=IFERROR(IF(E{fila}="","",E{fila}/(1-H{fila})),'
                   f'"revisa merma")')
        _reg(ws, f'I{fila}', i.value)
        i.number_format = FMT_CANT

        j = ws.cell(row=fila, column=10)
        j.value = (f'=IF(G{fila}="?","revisa unidades",'
                   f'IFERROR(I{fila}*D{fila}/G{fila},""))')
        _reg(ws, f'J{fila}', j.value)
        j.number_format = FMT_EUR

    _cf_factor(ws, d0, d1)
    rehacer_dv(ws, Layout(ws), cat_rango)

    # ---- bloque de resultado (§1.2 + §1.4) -------------------------------
    _reconstruir_cola(ws, d0, d1, r_tot, informe)
    return Layout(ws)


def _cf_factor(ws, d0, d1):
    """Rojo sobre la columna «Factor» cuando la pareja de unidades no existe.

    Idempotente: borra primero cualquier regla previa sobre una columna G.
    """
    try:
        reglas = ws.conditional_formatting._cf_rules
    except AttributeError:
        reglas = None
    if reglas is not None:
        for rango in list(reglas.keys()):
            if str(rango.sqref).startswith('G'):
                del reglas[rango]
    ws.conditional_formatting.add(
        f'G{d0}:G{d1}',
        CellIsRule(operator='equal', formula=['"?"'],
                   fill=PatternFill('solid', start_color=ROJO_BG,
                                    end_color=ROJO_BG),
                   font=Font(color=ROJO_FG, bold=True)))


def _estilo(ws, fila, col):
    return copy.copy(ws.cell(row=fila, column=col)._style)


def _banda(ws, fila):
    """Estilos de las 10 columnas de una fila (plantilla de banda)."""
    return [copy.copy(ws.cell(row=fila, column=c)._style) for c in range(1, 11)]


def _aplicar_banda(ws, fila, banda):
    for c in range(1, 11):
        ws.cell(row=fila, column=c)._style = copy.copy(banda[c - 1])


# Bloque de resultado v2: (offset, etiqueta, banda, hay_entrada, valor_entrada,
#                          fmt_entrada, formula_J, fmt_J, clave)
COLA = [
    (0, CAB_TOTAL, 'tot', False, None, None, 'SUM', FMT_EUR, 'ingredientes'),
    (1, 'Coste elaboración (%)', 'in', True, 0.10, FMT_PCT,
     '=J{ingredientes}*I{elaboracion}', FMT_EUR, 'elaboracion'),
    (2, 'COSTE TOTAL DEL PLATO', 'oro', False, None, None,
     '=J{ingredientes}+J{elaboracion}', FMT_EUR, 'plato'),
    (3, ETIQ_RACIONES, 'in', True, 1, FMT_ENT, None, None, 'raciones'),
    (4, ETIQ_COSTE_RACION, 'oro', False, None, None,
     '=IFERROR(J{plato}/I{raciones},"")', FMT_EUR, 'coste_racion'),
    (5, None, 'vacia', False, None, None, None, None, None),
    (6, ETIQ_FC, 'fc', True, 0.30, FMT_PCT, None, None, 'fc_objetivo'),
    (7, 'PVP SUGERIDO (sin IVA)', 'pln', False, None, None,
     '=IFERROR(J{coste_racion}/I{fc_objetivo},"")', FMT_EUR, 'pvp_sin'),
    (8, ETIQ_IVA, 'in', True, 0.10, FMT_PCT, None, None, 'iva'),
    (9, 'PVP CON IVA', 'oro', False, None, None,
     '=IFERROR(J{pvp_sin}*(1+I{iva}),"")', FMT_EUR, 'pvp_con'),
    (10, 'Margen bruto (€)', 'pln', False, None, None,
     '=IFERROR(J{pvp_sin}-J{coste_racion},"")', FMT_EUR, 'margen_eur'),
    (11, 'Margen bruto (%)', 'pct', False, None, None,
     '=IFERROR((J{pvp_sin}-J{coste_racion})/J{pvp_sin},"")', FMT_PCT1,
     'margen_pct'),
    (12, None, 'vacia', False, None, None, None, None, None),
    (13, ETIQ_PVP_ACTUAL, 'in', True, None, FMT_EUR, None, None, 'pvp_actual'),
    (14, ETIQ_FC_REAL, 'oro', False, None, None,
     '=IFERROR(J{coste_racion}/I{pvp_actual},"")', FMT_PCT1, 'fc_real'),
]


def _reconstruir_cola(ws, d0, d1, r_tot, informe):
    """Reescribe entero el bloque de totales: 9 filas v1 -> 15 filas v2.

    Escritura ABSOLUTA: se borra la banda completa y se vuelve a pintar desde
    las plantillas de estilo del propio fichero, así que la 2.ª pasada produce
    exactamente lo mismo.
    """
    # Valores del cliente que hay que conservar entre pasadas (§: el FC
    # objetivo de cada plantilla NO es el mismo — 0,20 en cócteles, 0,35 en
    # catering — y sobrescribirlo con el genérico sería un cambio de producto).
    # La clave se normaliza cortando por ' — ': en la 2.ª pasada la etiqueta del
    # food cost puede llevar el sufijo «— se edita en el Resumen» que le pone
    # grupo_a, y buscarla literal devolvía None → el objetivo se reseteaba.
    previos = {}
    for r in range(r_tot, ws.max_row + 1):
        etiqueta = ws.cell(row=r, column=1).value
        if isinstance(etiqueta, str) and etiqueta.strip():
            previos[_canon(etiqueta)] = ws.cell(row=r, column=9).value
    fc_previo = previos.get(ETIQ_FC)
    iva_previo = previos.get(ETIQ_IVA)
    rac_previo = previos.get(ETIQ_RACIONES)
    ela_previo = previos.get('Coste elaboración (%)')

    # Las plantillas de estilo se buscan por ETIQUETA, no por desplazamiento:
    # la cola de v1 tiene 9 filas y la de v2 quince, así que en la 2.ª pasada
    # r_tot+4 ya no es «Food Cost objetivo» sino «COSTE POR RACIÓN» y la
    # etiqueta del objetivo se pintaba de oro (28 diferencias en la primera
    # prueba de idempotencia).
    def _fila_etq(prefijo, defecto):
        for r in range(r_tot, min(ws.max_row, r_tot + 20) + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and _canon(v).startswith(prefijo):
                return r
        return defecto

    def _fila_vacia(defecto):
        for r in range(r_tot + 3, min(ws.max_row, r_tot + 20) + 1):
            if ws.cell(row=r, column=1).value is None:
                return r
        return defecto

    bandas = {
        'tot': _banda(ws, _fila_etq(CAB_TOTAL, r_tot)),
        'in': _banda(ws, _fila_etq('Coste elaboración', r_tot + 1)),
        'oro': _banda(ws, _fila_etq('COSTE TOTAL DEL PLATO', r_tot + 2)),
        'vacia': _banda(ws, _fila_vacia(r_tot + 3)),
        'fc': _banda(ws, _fila_etq('Food Cost objetivo', r_tot + 4)),
        'pln': _banda(ws, _fila_etq('PVP SUGERIDO', r_tot + 5)),
        'pct': _banda(ws, _fila_etq('Margen bruto (%)', r_tot + 8)),
    }

    fin_viejo = max(r_tot + 14, ws.max_row)
    for m in [x for x in ws.merged_cells.ranges
              if x.min_row >= r_tot and x.min_col == 1]:
        ws.unmerge_cells(str(m))
    for fila in range(r_tot, fin_viejo + 1):
        for col in range(1, 11):
            ws.cell(row=fila, column=col).value = None

    R = {clave: r_tot + off for off, _, _, _, _, _, _, _, clave in COLA if clave}
    R['_d0'], R['_d1'] = d0, d1
    defectos = {'elaboracion': ela_previo if isinstance(ela_previo, (int, float)) else 0.10,
                'raciones': rac_previo if isinstance(rac_previo, (int, float)) else 1,
                'fc_objetivo': fc_previo if isinstance(fc_previo, (int, float)) else 0.30,
                'iva': iva_previo if isinstance(iva_previo, (int, float)) else 0.10}

    for off, etiqueta, banda, hay_in, valor_in, fmt_in, formula, fmt_j, clave in COLA:
        fila = r_tot + off
        _aplicar_banda(ws, fila, bandas[banda])
        if etiqueta is None:
            continue
        ws.cell(row=fila, column=1, value=etiqueta)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                       end_column=8 if hay_in else 9)
        if hay_in:
            ci = ws.cell(row=fila, column=9)
            v = defectos.get(clave, valor_in)
            if v is not None:
                ci.value = v
            ci.number_format = fmt_in or FMT_PCT
            _pintar(ci, VERDE, ci.number_format, 'center', locked=False)
        if formula:
            if formula == 'SUM':
                formula = f'=SUM(J{d0}:J{d1})'
            else:
                formula = formula.format(**R)
            cv = ws.cell(row=fila, column=10, value=formula)
            cv.number_format = fmt_j
            _reg(ws, f'J{fila}', formula)

    _limpiar_cf(ws, R['fc_real'])
    ws.conditional_formatting.add(
        f"J{R['fc_real']}",
        FormulaRule(formula=[f"AND(ISNUMBER($J${R['fc_real']}),"
                             f"$J${R['fc_real']}>$I${R['fc_objetivo']})"],
                    fill=PatternFill('solid', start_color=ROJO_BG,
                                     end_color=ROJO_BG),
                    font=Font(color=ROJO_FG, bold=True), stopIfTrue=False))
    informe.append(f"{ws.title}: bloque raciones/food cost real en filas "
                   f"{R['raciones']}-{R['fc_real']}")
    return R


def _limpiar_cf(ws, fila_fcr):
    """Quita reglas de formato condicional previas sobre la celda del FC real
    (idempotencia: sin esto la 2.ª pasada apila una regla igual)."""
    try:
        reglas = ws.conditional_formatting._cf_rules
    except AttributeError:
        return
    for rango in list(reglas.keys()):
        if str(rango.sqref) == f'J{fila_fcr}':
            del reglas[rango]


# ==========================================================================
# §1.8 — Instrucciones
# ==========================================================================
def reescribir_instrucciones(wb, fname, informe):
    if 'Instrucciones' not in wb.sheetnames:
        return
    ws = wb['Instrucciones']
    col = 2 if ws.cell(row=2, column=2).value else 1
    modelo_titulo = copy.copy(ws.cell(row=6, column=col)._style)
    modelo_linea = copy.copy(ws.cell(row=8, column=col)._style)

    for r in range(6, ws.max_row + 1):
        ws.cell(row=r, column=col).value = None

    fila = 6
    def titulo(t):
        nonlocal fila
        c = ws.cell(row=fila, column=col, value=t)
        c._style = copy.copy(modelo_titulo)
        fila += 1

    def linea(t):
        nonlocal fila
        c = ws.cell(row=fila, column=col, value=t)
        c._style = copy.copy(modelo_linea)
        fila += 1

    titulo(TITULO_COMO_USAR)
    fila += 1
    for t in PROPIAS.get(fname, []):
        linea(t)
    linea('▸ Las celdas VERDES son las editables; el resto son fórmulas.')
    fila += 1
    override = BLOQUE_OVERRIDE.get(fname, {})
    for tit, lineas in BLOQUE_COLUMNAS:
        titulo(tit)
        fila += 1
        cuerpo = override.get(tit, lineas)
        if cuerpo == '@mermas':                      # TEC-R2-09
            cuerpo = _lineas_mermas()
        for t in cuerpo:
            linea(_sustantivo(fname, t))             # COM-B16
        fila += 1
    linea(PIE)
    fila += 1
    linea(COPY)
    fila += 1
    linea(VERSION_LINE)
    informe.append('Instrucciones reescritas (§1.8)')


# ==========================================================================
# Referencias CRUZADAS entre hojas
# ==========================================================================
# insertar_columna/insertar_fila sólo traducen las fórmulas de SU hoja. Las que
# apuntan desde otra pestaña ('Presupuesto'!C6 = ='Cocktail 50 pax'!I23) se
# quedan clavadas en la coordenada vieja y pasan a leer una celda que ya no es
# el total — el 06 se quedó sin coste por persona en la primera prueba, y el
# censo lo cazó como nocache_real. Aquí se reescriben una a una.
def _cola_v1(r_tot):
    """Direcciones de los resultados en el layout v1 (antes de tocar nada)."""
    return ['I%d' % r_tot, 'H%d' % (r_tot + 1), 'I%d' % (r_tot + 1),
            'I%d' % (r_tot + 2), 'H%d' % (r_tot + 4), 'I%d' % (r_tot + 5),
            'I%d' % (r_tot + 6), 'I%d' % (r_tot + 7), 'I%d' % (r_tot + 8)]


def _cola_v2(b):
    return ['J%d' % b['ingredientes'], 'I%d' % b['elaboracion'],
            'J%d' % b['elaboracion'], 'J%d' % b['plato'],
            'I%d' % b['fc_objetivo'], 'J%d' % b['pvp_sin'],
            'J%d' % b['pvp_con'], 'J%d' % b['margen_eur'],
            'J%d' % b['margen_pct']]


def remapear_referencias(wb, mapa, informe):
    """mapa = {hoja_destino: {ref_vieja: ref_nueva}}."""
    n = 0
    destinos = set(mapa)
    for ws in wb.worksheets:
        if ws.title in destinos:
            continue          # dentro de la hoja ya lo hizo insertar_*
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith('=')):
                    continue
                nuevo = v
                for hoja, refs in mapa.items():
                    if hoja not in nuevo:
                        continue
                    for viejo, nuevo_ref in refs.items():
                        col, fila = viejo[0], viejo[1:]
                        patron = re.compile(
                            r"((?:'" + re.escape(hoja) + r"'|" +
                            re.escape(hoja) + r")!\$?)" + col +
                            r"(\$?)" + fila + r"(?![0-9])")
                        nuevo = patron.sub(
                            lambda m, nr=nuevo_ref: m.group(1) + nr[0] +
                            m.group(2) + nr[1:], nuevo)
                if nuevo != v:
                    c.value = nuevo
                    n += 1
                    informe.append(f'{ws.title}!{c.coordinate}: referencia '
                                   f'cruzada actualizada → {nuevo}')
    return n


# ==========================================================================
# API principal
# ==========================================================================
def aplicar(wb, fname, informe):
    """Aplica el motor §1 a todas las hojas de escandallo del libro."""
    _, cf0, cf1 = escribir_conversiones(wb, informe)
    _, mf0, mf1 = escribir_mermas(wb, informe)
    conv_rango = f'Conversiones!$A${cf0}:$B${cf1}'
    mermas_rango = f'Mermas!$A${mf0}:$C${mf1}'
    cat_rango = f'Mermas!$A${mf0}:$A${mf1}'

    pendientes = {ws.title: _cola_v1(lay.r_tot)
                  for ws, lay in hojas_escandallo(wb) if not lay.v2}

    layouts = {}
    for ws, _ in hojas_escandallo(wb):
        lay = aplicar_hoja(ws, fname, informe, mermas_rango, conv_rango,
                           cat_rango)
        if lay:
            layouts[ws.title] = lay

    if pendientes:
        mapa = {t: dict(zip(viejas, _cola_v2(bloque(wb[t]))))
                for t, viejas in pendientes.items()}
        remapear_referencias(wb, mapa, informe)

    reescribir_instrucciones(wb, fname, informe)
    return layouts


def bloque(ws):
    """Devuelve {clave: fila} del bloque de resultado de una hoja ya en v2."""
    etiquetas = {
        CAB_TOTAL: 'ingredientes', 'Coste elaboración (%)': 'elaboracion',
        'COSTE TOTAL DEL PLATO': 'plato', ETIQ_RACIONES: 'raciones',
        ETIQ_COSTE_RACION: 'coste_racion', ETIQ_FC: 'fc_objetivo',
        'PVP SUGERIDO (sin IVA)': 'pvp_sin', ETIQ_IVA: 'iva',
        'PVP CON IVA': 'pvp_con', 'Margen bruto (€)': 'margen_eur',
        'Margen bruto (%)': 'margen_pct', ETIQ_PVP_ACTUAL: 'pvp_actual',
        ETIQ_FC_REAL: 'fc_real',
    }
    fuera = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            # grupo_a añade « — se edita en el Resumen» a la etiqueta del food
            # cost objetivo y grupo_b renombra rótulos por contexto; sin
            # normalizar, bloque() dejaba de encontrarlos.
            clave = _canon(v)
            if clave in etiquetas:
                fuera[etiquetas[clave]] = r
    return fuera


# ==========================================================================
# §1.6 — Protección + cierre del libro
# ==========================================================================
def _es_verde(cel):
    f = cel.fill
    return (f is not None and f.fill_type == 'solid' and f.fgColor is not None
            and isinstance(f.fgColor.rgb, str)
            and f.fgColor.rgb.upper().endswith(VERDE))


def proteger(ws, informe):
    """Protección sin contraseña: se desbloquean SOLO las celdas verdes."""
    verdes = 0
    for row in ws.iter_rows():
        for c in row:
            if _es_verde(c):
                c.protection = Protection(locked=False)
                verdes += 1
            else:
                c.protection = Protection(locked=True)
    ws.protection.sheet = True
    # SIN contraseña (SPEC §6). Ojo: NO tocar `password`. `= None` revienta
    # openpyxl y `= ''` escribe el hash de la cadena vacía (CE4B) → Excel pide
    # contraseña al desproteger, justo lo contrario de lo que dicen las
    # Instrucciones. Dejándolo sin asignar, el atributo no se escribe.
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    informe.append(f'{ws.title}: protegida sin contraseña ({verdes} celdas '
                   'verdes editables)')
    return verdes


def ajustar_notas(ws):
    """TEC-R2-03: los avisos de la fila 2 van COMBINADOS (A2:J2) y sin
    `wrap_text`, así que Excel los corta por el ancho de la combinación en vez
    de repartirlos en varias líneas: el AVISO de 06!'Cocktail (por persona)'
    —174 caracteres sobre 137 de ancho— se leía a medias en pantalla y en
    papel. Se activa el ajuste de texto y se da altura a la fila."""
    for m in list(ws.merged_cells.ranges):
        if m.min_col != 1 or m.min_row != m.max_row or not 2 <= m.min_row <= 3:
            continue
        c = ws.cell(row=m.min_row, column=1)
        if not isinstance(c.value, str) or len(c.value) < 70:
            continue
        ancho = 0.0
        for i in range(m.min_col, m.max_col + 1):
            ancho += ws.column_dimensions[get_column_letter(i)].width or 8.43
        c.alignment = Alignment(wrap_text=True, vertical='top',
                                horizontal=c.alignment.horizontal)
        lineas = max(1, -(-len(c.value) // max(20, int(ancho) - 2)))
        ws.row_dimensions[m.min_row].height = max(15.0, 13.0 * lineas + 4)


def _altura_foto(ws, cel):
    """El rótulo de la foto pasa de 2 a 3 líneas: la combinación L8:M9 con
    altura por defecto (2 × 15 pt) lo cortaría."""
    for m in ws.merged_cells.ranges:
        if m.min_row <= cel.row <= m.max_row and m.min_col <= cel.column <= m.max_col:
            for r in range(m.min_row, m.max_row + 1):
                ws.row_dimensions[r].height = max(
                    ws.row_dimensions[r].height or 15.0, 22.0)
            return


FILAS_FOTO = 11          # cabecera + 10: el bloque L4:M14
ALTO_FOTO = 28.0         # pt


def _altura_zona_foto(ws, lay):
    """TEC-R2-13 (R1 TEC-29): la «zona para foto del plato» que anuncia el hero
    de la landing eran DOS filas de 15 pt — 1 cm de alto por 7,4 de ancho, una
    ranura, no un hueco para una foto. Con 11 filas a 28 pt el bloque L4:M14
    pasa a ~7,4 × 10,8 cm, proporción de foto de plato, y las primeras filas de
    ingredientes ganan aire. `print_setup` deja `fitToHeight=0`, así que crecer
    a lo alto no rompe el A4."""
    hay = any(ws.cell(row=r, column=c).value
              for r in range(max(1, lay.hdr - 1), lay.hdr + FILAS_FOTO)
              for c in (12, 13))
    if not hay:
        return                       # esta hoja no tiene zona de foto
    for r in range(lay.hdr, lay.hdr + FILAS_FOTO):
        actual = ws.row_dimensions[r].height
        if actual is None or actual < ALTO_FOTO:
            ws.row_dimensions[r].height = ALTO_FOTO


def cerrar(wb, fname, informe, proteger_hojas=True):
    """Cierre común: limpieza de '', foto, impresión A4, protección y versión."""
    for ws in wb.worksheets:
        ajustar_notas(ws)
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    c.value = None
                elif (isinstance(c.value, str)
                      and c.value.startswith('Arrastra aquí la foto')):
                    c.value = _texto_foto(c.value)      # DOM-14
                    _altura_foto(ws, c)
                elif (isinstance(c.value, str)
                      and c.value.startswith('Insertar → Imagen')):
                    c.value = None                      # DOM-14: sobra
        if ws.title == 'Instrucciones':
            if ws.page_setup.paperSize != 9:
                print_setup(ws, None, landscape=False)
            else:
                ws.page_setup.paperSize = 9
            linea_instrucciones(ws, VERSION_LINE, RX_VERSION)
            continue
        lay = Layout(ws)
        if lay.valida:
            _altura_zona_foto(ws, lay)               # TEC-R2-13
        print_setup(ws, lay.hdr if lay.valida else None,
                    landscape=ws.max_column >= 6)
        ultima_col = get_column_letter(max(10, ws.max_column))
        if lay.valida:
            ultima_col = 'J'
        ws.print_area = f'A1:{ultima_col}{ws.max_row}'
        if proteger_hojas and ws.title not in ('Conversiones', 'Mermas'):
            proteger(ws, informe)
        elif ws.title in ('Conversiones', 'Mermas'):
            # auxiliares: el cliente TIENE que poder añadir parejas y ajustar
            # mermas, así que se quedan sin protección (y así lo dicen las
            # Instrucciones).
            ws.protection.sheet = False
