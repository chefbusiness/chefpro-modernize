#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 del Kit de Escandallos Pro.

    python3 main.py --dry-run [--solo a|b|c] [--json informe.json]

REGLA DURA: `astro-site/public/dl/kit-escandallos/` NO se toca. `--dry-run`
regenera una copia en el scratchpad y trabaja allí. Sin `--dry-run` el script
ABORTA salvo que se le pase `KIT_ESCANDALLOS_APPLY=1` en el entorno: la
ejecución real la hace el orquestador cuando la ronda 2 dé verde (SPEC, cabecera).

Qué hace, en este orden:
  1. Copia de trabajo (dry-run) desde `dl/kit-escandallos`.
  1b. BONUS 1 (`bono_guia.py`) desde el .md del repo: el .pdf va DENTRO del
     build —la copia de trabajo se regenera en cada pasada y antes se llevaba
     el bono por delante (DOM-06)— y el .docx a `_bono/` del scratchpad, fuera
     de la carpeta de entrega (COM-B18). Se comprueba que el PDF no tiene
     ningún ■ (COM-A01).
  2. Por fichero: grupo.pre → motor.aplicar (§1) → grupo.post (§2) → motor.cerrar.
  3. Idempotencia: repite el pipeline sobre un clon del resultado y compara
     celda a celda. Debe dar 0 diferencias.
  4. `inject_cache.py` sobre los ficheros tocados (SIEMPRE al final: cualquier
     guardado posterior de openpyxl borraría el cache).
  5. Verificación `data_only` de todas las fórmulas nuevas del registro.
  6. Sensibilidad con pycel: 2 inputs por fichero; el coste tiene que moverse.
  7. `censo-entregables.py --only <carpeta> --fail`.
  8. Demostraciones para el informe (sobre copias desechables, nunca sobre los
     entregables): factor en las 17 filas de unidad mixta, IVA en celda, merma
     por categoría al elegir «Pescado» en una fila vacía, 9 pases del 02,
     rotación semanal del 03, carrot cake del 07 y protección.

Térmica: todo en SERIE. No hay builds ni navegador.
"""
import argparse
import contextlib
import datetime
import importlib
import json
import logging
import os
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402
from openpyxl.utils import get_column_letter               # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
ORIGEN = os.path.join(ROOT, 'astro-site', 'public', 'dl', 'kit-escandallos')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    'f83a04d3-d14f-49b3-aa79-c442fb4d7983/scratchpad')
DESTINO = os.path.join(SCRATCH, 'dryrun-v2', 'kit-escandallos')
IDEM = os.path.join(SCRATCH, 'dryrun-v2', 'kit-escandallos-idem')
DEMOS = os.path.join(SCRATCH, 'dryrun-v2', '_demos')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
# DOM-06: el BONUS 1 lo producía un script suelto que NO estaba encadenado, así
# que cada `--dry-run` regeneraba la carpeta desde `dl/` y BORRABA el .pdf y el
# .docx del bono. Con la clave `bonus-guia` ya dada de alta en PRODUCT_FILES y
# la tarjeta ya en el dashboard, un despliegue tras un rebuild limpio dejaba al
# cliente pulsando un botón que da 404. Ahora el bono es un paso más del build.
BONO = os.path.join(AQUI, 'bono_guia.py')
BONO_MD = os.path.join(AQUI, 'bono-guia-food-cost-30-dias.md')
BONO_BASE = 'BONUS-guia-food-cost-30-dias'
# COM-B18: el .docx NO es un entregable (no tiene clave en PRODUCT_FILES ni
# tarjeta en el dashboard). Se escribe fuera de la carpeta de entrega.
BONO_DOCX_DIR = os.path.join(SCRATCH, 'dryrun-v2', '_bono')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')

SPEC = 'scripts/productos-digitales/kit-escandallos-v2-SPEC.md §1 y §2'


# ==========================================================================
# Utilidades
# ==========================================================================
def log(msg):
    print(msg, flush=True)


def preparar_copia():
    if not os.path.isdir(ORIGEN):
        raise SystemExit(f'No existe el origen: {ORIGEN}')
    if os.path.isdir(DESTINO):
        shutil.rmtree(DESTINO)
    os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
    shutil.copytree(ORIGEN, DESTINO)
    log(f'  copia de trabajo regenerada: {DESTINO}')


def construir_bono(carpeta):
    """Regenera el BONUS 1: el .pdf en la carpeta de entregables y el .docx
    FUERA (COM-B18).

    El .docx no tiene clave en `PRODUCT_FILES` ni tarjeta en el dashboard: si
    se queda en `/dl/kit-escandallos` es un fichero huérfano que nadie puede
    descargar y que el censo cuenta como entregable. Se deja en `_bono/` del
    scratchpad, junto al build, para poder abrirlo y editarlo.
    """
    salida = os.path.join(carpeta, BONO_BASE)
    os.makedirs(BONO_DOCX_DIR, exist_ok=True)
    salida_docx = os.path.join(BONO_DOCX_DIR, BONO_BASE)
    r = subprocess.run([sys.executable, BONO, BONO_MD, salida, salida_docx],
                       capture_output=True, text=True)
    log('    ' + (r.stdout.strip() or r.stderr.strip()[-300:]))
    hay = {'.pdf': os.path.isfile(salida + '.pdf'),
           '.docx': os.path.isfile(salida_docx + '.docx')}
    rutas = {'.pdf': salida + '.pdf', '.docx': salida_docx + '.docx'}
    fallos = []
    if r.returncode != 0:
        fallos.append(f'bono_guia.py devolvió {r.returncode}')
    for ext, ok in hay.items():
        if not ok:
            fallos.append(f'el bono no generó {BONO_BASE}{ext}')
    # COM-B18: y que NO se cuele el .docx en la carpeta de entrega.
    if os.path.isfile(salida + '.docx'):
        fallos.append(f'el .docx sigue en la carpeta de entrega: '
                      f'{salida}.docx (COM-B18)')
    fuera = {'exit': r.returncode, 'ficheros': hay, 'rutas': rutas,
             'bytes': {e: (os.path.getsize(rutas[e]) if ok else 0)
                       for e, ok in hay.items()},
             'fallos': fallos}
    fuera['pdf_sin_cuadrados'] = _pdf_sin_cuadrados(rutas['.pdf'], fallos)
    return fuera


def _pdf_sin_cuadrados(path, fallos):
    """COM-A01: extrae el texto del PDF y comprueba que no queda ningún ■ ni
    ningún carácter fuera de WinAnsi, y que las tildes, la ñ, el € y las
    comillas latinas SÍ se ven."""
    if not os.path.isfile(path):
        return {'ok': False, 'motivo': 'no hay PDF'}
    texto, via = '', None
    for mod, attr in (('pypdf', 'PdfReader'), ('PyPDF2', 'PdfReader')):
        try:
            lector = getattr(importlib.import_module(mod), attr)(path)
            texto = '\n'.join((pg.extract_text() or '') for pg in lector.pages)
            via = mod
            break
        except Exception:                                        # noqa: BLE001
            continue
    if via is None:
        return {'ok': None, 'motivo': 'ni pypdf ni PyPDF2 instalados'}
    malos = {}
    for ch in texto:
        if ch in '\n\r\t':
            continue
        try:
            ch.encode('cp1252')
        except UnicodeEncodeError:
            malos[ch] = malos.get(ch, 0) + 1
    cuadrados = sum(texto.count(c) for c in '\u25a0\u25aa\ufffd')
    esperados = {'tildes': any(c in texto for c in 'áéíóúÁÉÍÓÚ'),
                 'enye': 'ñ' in texto or 'Ñ' in texto,
                 'euro': '€' in texto,
                 'comillas': '«' in texto and '»' in texto}
    if cuadrados:
        fallos.append(f'el PDF del bono tiene {cuadrados} cuadrados ■')
    if malos:
        fallos.append('el PDF del bono tiene caracteres fuera de WinAnsi: '
                      + ', '.join(f'{c!r}×{n}' for c, n in malos.items()))
    for k, ok in esperados.items():
        if not ok:
            fallos.append(f'el PDF del bono no muestra {k}')
    return {'ok': not cuadrados and not malos and all(esperados.values()),
            'via': via, 'caracteres': len(texto), 'cuadrados': cuadrados,
            'fuera_de_winansi': {c: n for c, n in malos.items()},
            'se_ven': esperados}


def digest(path):
    """Huella comparable de un .xlsx (valores, formatos, relleno, merges, DV)."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked))
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted(f'{dv.type}:{dv.formula1}:{dv.sqref}'
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted(str(r.sqref) for r in ws.conditional_formatting),
            'prot': bool(ws.protection.sheet),
            'area': str(ws.print_area),
        }
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append(f'{fichero}:{hoja}: hoja sólo en una pasada')
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'prot', 'area'):
            if ha[k] != hb[k]:
                fuera.append(f'{fichero}:{hoja}: cambia {k}')
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append(f'{fichero}:{hoja}!{coord}: '
                             f'{ca.get(coord)} → {cb.get(coord)}')
    return fuera


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, grupos, informe_global):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    cambios, registro_grupo = [], []
    motor.REGISTRO = []

    # Un grupo puede declarar `PROPIOS`: ficheros que NO llevan rejilla de
    # escandallo (09, 10, 11, BONUS). Ahí el motor §1 no aplica — les colgaría
    # las hojas «Conversiones»/«Mermas» y les reescribiría las Instrucciones
    # describiendo columnas que no tienen —, así que el grupo hace el trabajo
    # entero y llama él mismo a `motor.cerrar`.
    propio = any(fname in getattr(g, 'PROPIOS', []) for g in grupos)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []):
            g.pre(wb, fname, cambios)

    if not propio:
        motor.aplicar(wb, fname, cambios)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []):
            g.post(wb, fname, cambios, registro_grupo)

    if not propio:
        motor.cerrar(wb, fname, cambios)
    wb.save(path)

    registro = [(h, c, f) for h, c, f in motor.REGISTRO]
    registro += [(h, c, f) for h, c, f in registro_grupo
                 if isinstance(f, str) and f.startswith('=')]
    informe_global.append({'fichero': fname, 'cambios': cambios,
                           'formulas_nuevas': len(registro)})
    return registro


def ficheros_a_tocar(grupos):
    nombres = list(motor.FICHEROS)
    for g in grupos:
        for f in getattr(g, 'FICHEROS', []):
            if f not in nombres:
                nombres.append(f)
    return nombres


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        fuera.append((n, r.stdout.strip(), r.returncode))
        log('    ' + r.stdout.strip())
        if r.returncode != 0:
            log('    ERROR inject_cache: ' + r.stderr[-400:])
    return fuera


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA debe tener valor cacheado, salvo las que devuelven ""
    por diseño (su texto contiene una cadena vacía)."""
    fallos, vacias, ok = [], 0, 0
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        wbv = openpyxl.load_workbook(path, data_only=True)
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append(f'{fname}:{hoja}!{coord}: hoja ausente')
                continue
            v = wbv[hoja][coord].value
            if v is None:
                if '""' in formula:
                    vacias += 1
                else:
                    fallos.append(f'{fname}:{hoja}!{coord}: sin cache '
                                  f'({formula[:60]})')
            else:
                ok += 1
    return {'con_valor': ok, 'vacias_por_diseno': vacias, 'fallos': fallos}


def sensibilidad(carpeta, nombres):
    """pycel: cambia 2 inputs por fichero; el coste TIENE que moverse."""
    from pycel import ExcelCompiler
    fuera, fallos = [], []
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        wb = openpyxl.load_workbook(path)
        hoja = None
        for ws, lay in motor.hojas_escandallo(wb):
            if ws.cell(row=lay.d0, column=1).value:
                hoja, layout = ws.title, lay
                break
        if hoja is None:
            continue
        b = motor.bloque(wb[hoja])
        destino = f"'{hoja}'!J{b['plato']}"
        d0 = layout.d0
        xl = ExcelCompiler(filename=path)
        with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
            base = xl.evaluate(destino)
            pruebas = []
            for col, factor in (('E', 2.0), ('D', 3.0)):
                ref = f"'{hoja}'!{col}{d0}"
                previo = xl.evaluate(ref)
                xl.set_value(ref, (previo or 1) * factor)
                nuevo = xl.evaluate(destino)
                pruebas.append({'input': f'{hoja}!{col}{d0}',
                                'de': previo, 'a': (previo or 1) * factor,
                                'coste_antes': round(base, 4),
                                'coste_despues': round(nuevo, 4),
                                'cambia': abs(nuevo - base) > 1e-9})
                xl.set_value(ref, previo)
                base = xl.evaluate(destino)
        for p in pruebas:
            if not p['cambia']:
                fallos.append(f"{fname}: {p['input']} no mueve el coste")
        fuera.append({'fichero': fname, 'celda_coste': destino,
                      'pruebas': pruebas})
    return fuera, fallos


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-1500:])
    return {'exit': r.returncode, 'salida': r.stdout.strip().splitlines()[-6:]}


# ==========================================================================
# Demostraciones para el informe
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}'


def demo_factor(carpeta, origen=None):
    """Las 17 filas con Ud. Compra ≠ Ud. Uso: coste antes (fichero original
    v1.1: en producción es el RESPALDO, porque ORIGEN ya está sobrescrito)
    y después (fichero v2, valor cacheado)."""
    origen = origen or ORIGEN
    fuera = []
    por_fichero = {}
    for fname, hoja, fila in motor.FILAS_UNIDAD_MIXTA:
        por_fichero.setdefault(fname, []).append((hoja, fila))
    for fname, filas in por_fichero.items():
        wb0 = openpyxl.load_workbook(os.path.join(origen, fname))
        wbv = openpyxl.load_workbook(os.path.join(carpeta, fname),
                                     data_only=True)
        wbf = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for hoja, fila in filas:
            o = wb0[hoja]
            d, e, g = (o.cell(row=fila, column=c).value for c in (4, 5, 7))
            coste_antes = e / (1 - (g or 0)) * d
            v = wbv[hoja]
            f = wbf[hoja]
            fuera.append({
                'ref': f'{fname}:{hoja}:{fila}',
                'ingrediente': f.cell(row=fila, column=1).value,
                'antes': {'ud_compra': o.cell(row=fila, column=3).value,
                          'cantidad': e,
                          'ud_uso': o.cell(row=fila, column=6).value,
                          'coste_eur': round(coste_antes, 4)},
                'despues': {'ud_compra': f.cell(row=fila, column=3).value,
                            'cantidad': f.cell(row=fila, column=5).value,
                            'ud_uso': f.cell(row=fila, column=6).value,
                            'factor': v.cell(row=fila, column=7).value,
                            'coste_eur': round(v.cell(row=fila, column=10).value, 4)
                            if isinstance(v.cell(row=fila, column=10).value,
                                          (int, float)) else None},
                'motivo': motor.PARCHES.get((fname, hoja, fila), ({}, ''))[1],
            })
    return fuera


def demo_iva(carpeta, nombres):
    fuera = []
    for fname in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        wbv = openpyxl.load_workbook(os.path.join(carpeta, fname),
                                     data_only=True)
        for ws, _ in motor.hojas_escandallo(wb):
            b = motor.bloque(ws)
            if 'iva' not in b:
                continue
            fuera.append({
                'ref': f"{fname}:{ws.title}:I{b['iva']}",
                'etiqueta': ws.cell(row=b['iva'], column=1).value,
                'valor': ws.cell(row=b['iva'], column=9).value,
                'formula_pvp_con_iva': ws.cell(row=b['pvp_con'], column=10).value,
                'pvp_con_iva_calculado': wbv[ws.title].cell(
                    row=b['pvp_con'], column=10).value,
            })
            break        # una hoja por fichero basta como muestra
    return fuera


def demo_merma(carpeta):
    """Fila VACÍA + categoría «Pescado» → la merma típica (35 %) entra sola.
    Se hace sobre una COPIA desechable: el entregable no se toca."""
    os.makedirs(DEMOS, exist_ok=True)
    src = os.path.join(carpeta, '01-escandallo-estandar.xlsx')
    dst = os.path.join(DEMOS, 'demo-merma-01.xlsx')
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb['Escandallo']
    lay = motor.Layout(ws)
    fila = None
    for r in range(lay.d0, lay.d1 + 1):
        if ws.cell(row=r, column=1).value is None:
            fila = r
            break
    ws.protection.sheet = False
    ws.cell(row=fila, column=1, value='Merluza (prueba)')
    ws.cell(row=fila, column=2, value='Pescado')
    ws.cell(row=fila, column=3, value='kg')
    ws.cell(row=fila, column=4, value=14)
    ws.cell(row=fila, column=5, value=200)
    ws.cell(row=fila, column=6, value='g')
    wb.save(dst)
    xl = _pycel(dst)
    return {
        'copia_desechable': dst,
        'fila_vacia': f'01-escandallo-estandar.xlsx:Escandallo:{fila}',
        'se_escribe': {'B': 'Pescado', 'C': 'kg', 'D': 14, 'E': 200, 'F': 'g'},
        'formula_merma': ws.cell(row=fila, column=8).value,
        'merma_resultante': _ev(xl, f"'Escandallo'!H{fila}"),
        'factor_resultante': _ev(xl, f"'Escandallo'!G{fila}"),
        'cant_bruta': _ev(xl, f"'Escandallo'!I{fila}"),
        'coste_eur': _ev(xl, f"'Escandallo'!J{fila}"),
        'lectura': ('200 g netos de merluza a 14 €/kg: la merma del 35 % sale '
                    'de la hoja «Mermas» al elegir la categoría y el factor '
                    'kg→g (1000) evita el ×1.000 que denunciaba DOM-05.'),
    }


def demo_02(carpeta):
    path = os.path.join(carpeta, '02-menu-degustacion.xlsx')
    wbv = openpyxl.load_workbook(path, data_only=True)
    wb = openpyxl.load_workbook(path)
    rs, rsv = wb['Resumen'], wbv['Resumen']
    pases = []
    for r in range(5, 14):
        pases.append({'pase': rs.cell(row=r, column=2).value,
                      'formula': rs.cell(row=r, column=3).value,
                      'valor': rsv.cell(row=r, column=3).value})
    fc = None
    for r in range(1, rs.max_row + 1):
        v = rs.cell(row=r, column=2).value
        if isinstance(v, str) and v.startswith('Food Cost objetivo'):
            fc = r
    # prueba viva: rellenar «6. Pase» en una copia desechable
    os.makedirs(DEMOS, exist_ok=True)
    dst = os.path.join(DEMOS, 'demo-02-pase6.xlsx')
    shutil.copy2(path, dst)
    wd = openpyxl.load_workbook(dst)
    ws6 = wd['6. Pase']
    ws6.protection.sheet = False
    lay = motor.Layout(ws6)
    ws6.cell(row=lay.d0, column=1, value='Bogavante (prueba)')
    ws6.cell(row=lay.d0, column=2, value='Marisco')
    ws6.cell(row=lay.d0, column=3, value='kg')
    ws6.cell(row=lay.d0, column=4, value=45)
    ws6.cell(row=lay.d0, column=5, value=0.12)
    ws6.cell(row=lay.d0, column=6, value='kg')
    wd.save(dst)
    xl = _pycel(dst)
    return {
        'hojas': wb.sheetnames,
        'resumen_pases': pases,
        'total_menu': rsv.cell(row=14, column=3).value,
        'food_cost_unico': {
            'celda': f'Resumen!C{fc}',
            'valor': rs.cell(row=fc, column=3).value,
            'lo_leen': [f"{t}: I{motor.bloque(wb[t])['fc_objetivo']} = "
                        f"{wb[t].cell(row=motor.bloque(wb[t])['fc_objetivo'], column=9).value}"
                        for t in wb.sheetnames if t.endswith('Pase') or t[0].isdigit()],
        },
        'prueba_pase_vacio_vs_lleno': {
            'copia_desechable': dst,
            'C10_antes': pases[5]['valor'],
            'C10_con_un_ingrediente': _ev(xl, "'Resumen'!C10"),
            'total_antes': rsv.cell(row=14, column=3).value,
            'total_despues': _ev(xl, "'Resumen'!C14"),
        },
    }


def demo_03(carpeta):
    path = os.path.join(carpeta, '03-menu-del-dia.xlsx')
    wb = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    ws, wsv = wb['Rotación Semanal'], wbv['Rotación Semanal']
    filas = []
    for r in range(5, 12):
        if ws.cell(row=r, column=1).value is None:
            continue
        filas.append({
            'dia': ws.cell(row=r, column=1).value,
            'coste_menu': wsv.cell(row=r, column=9).value,
            'formula_coste': ws.cell(row=r, column=9).value,
            'pvp_sin_iva': wsv.cell(row=r, column=10).value,
            'food_cost': wsv.cell(row=r, column=11).value,
        })
    return {'hojas': wb.sheetnames,
            'cabecera': [ws.cell(row=4, column=c).value for c in range(1, 12)],
            'filas': filas}


def demo_07(carpeta):
    path = os.path.join(carpeta, '07-cafeteria-brunch.xlsx')
    wb = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    ws, wsv = wb['Carrot Cake'], wbv['Carrot Cake']
    b = motor.bloque(ws)
    lay = motor.Layout(ws)
    ingredientes = []
    for r in range(lay.d0, lay.d1 + 1):
        if ws.cell(row=r, column=1).value is None:
            continue
        ingredientes.append({
            'fila': r, 'ingrediente': ws.cell(row=r, column=1).value,
            'coste': wsv.cell(row=r, column=10).value})
    return {
        'hojas': wb.sheetnames,
        'titulo': ws['A1'].value,
        'ingredientes': ingredientes,
        'raciones': ws.cell(row=b['raciones'], column=9).value,
        'coste_tanda': wsv.cell(row=b['plato'], column=10).value,
        'coste_racion': wsv.cell(row=b['coste_racion'], column=10).value,
        'pvp_sin_iva': wsv.cell(row=b['pvp_sin'], column=10).value,
        'pvp_con_iva': wsv.cell(row=b['pvp_con'], column=10).value,
    }


def demo_tostada(carpeta):
    path = os.path.join(carpeta, '07-cafeteria-brunch.xlsx')
    wbv = openpyxl.load_workbook(path, data_only=True)
    wb = openpyxl.load_workbook(path)
    ws, wsv = wb['Tostada Aguacate'], wbv['Tostada Aguacate']
    b = motor.bloque(ws)
    return {'coste_plato': wsv.cell(row=b['plato'], column=10).value,
            'pvp_sin_iva': wsv.cell(row=b['pvp_sin'], column=10).value,
            'pvp_con_iva': wsv.cell(row=b['pvp_con'], column=10).value}


def _hay_hoja(carpeta, fname, hoja):
    """Las demos de grupo A miran hojas que sólo existen si grupo_a corrió.
    Con `--solo c` los 8 escandallos SÍ se procesan (motor.FICHEROS), pero sin
    «6. Pase» / «Rotación Semanal» / «Carrot Cake»: sin esta guarda el paso 7
    reventaba con un KeyError."""
    path = os.path.join(carpeta, fname)
    if not os.path.isfile(path):
        return False
    return hoja in openpyxl.load_workbook(path, read_only=True).sheetnames


def demo_proteccion(carpeta, nombres):
    fuera = []
    for fname in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wb.worksheets:
            verdes = sum(1 for row in ws.iter_rows() for c in row
                         if motor._es_verde(c) and not c.protection.locked)
            bloqueadas_formula = sum(
                1 for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith('=')
                and c.protection.locked)
            fuera.append({'fichero': fname, 'hoja': ws.title,
                          'protegida': bool(ws.protection.sheet),
                          'con_contrasena': bool(ws.protection.password),
                          'verdes_editables': verdes,
                          'formulas_bloqueadas': bloqueadas_formula})
    return fuera


# ==========================================================================
def main():
    ap = argparse.ArgumentParser(description='Post-proceso v2.0 Kit de Escandallos')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    ap.add_argument('--solo', default='a', help='grupos a aplicar: a, b, c o a,b')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--sin-idempotencia', action='store_true')
    args = ap.parse_args()

    if not args.dry_run and os.environ.get('KIT_ESCANDALLOS_APPLY') != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en '
            f'{ORIGEN}, que la SPEC declara intocable hasta que la ronda 2 dé '
            'verde. Usa --dry-run (o KIT_ESCANDALLOS_APPLY=1 si eres el '
            'orquestador y ya tienes el visto bueno).')

    carpeta = DESTINO if args.dry_run else ORIGEN
    respaldo = None
    if args.dry_run:
        preparar_copia()
    else:
        # TEC-R2-14 (2): en producción se escribe IN PLACE sobre los
        # entregables. Si algo revienta a mitad del bucle de 12 ficheros, los
        # xlsx quedan a medio transformar y sólo los recupera `git checkout`.
        # El respaldo va al SCRATCHPAD, nunca junto a los entregables: una
        # carpeta `.bak` dentro de `public/dl/` se publicaría.
        respaldo = os.path.join(
            SCRATCH, 'kit-escandallos.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(ORIGEN, respaldo)
        log(f'  respaldo previo de los entregables: {respaldo}')

    grupos = []
    pedidos = [g.strip().lower() for g in args.solo.split(',') if g.strip()]
    for letra in pedidos:
        try:
            grupos.append(importlib.import_module(f'grupo_{letra}'))
            log(f'  grupo_{letra} cargado')
        except ImportError:
            log(f'  grupo_{letra} no existe todavía — se omite')

    nombres = ficheros_a_tocar(grupos)
    informe_ficheros, registros = [], {}
    log(f'\n== 1/7 · post-proceso de {len(nombres)} ficheros ==')
    for fname in nombres:
        registros[fname] = procesar(carpeta, fname, grupos, informe_ficheros)
        log(f'  {fname}: {len(registros[fname])} fórmulas nuevas')

    log('\n== 1b/7 · BONUS 1 «Guía: Controla tu Food Cost en 30 Días» ==')
    bono = construir_bono(carpeta)

    # ---- idempotencia ---------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 2/7 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(IDEM):
            shutil.rmtree(IDEM)
        shutil.copytree(carpeta, IDEM)
        antes = {n: digest(os.path.join(carpeta, n)) for n in nombres}
        for fname in nombres:
            procesar(IDEM, fname, grupos, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(IDEM, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:40]}
        log(f'  diferencias 1.ª vs 2.ª pasada: {len(difs)}')
        for d in difs[:10]:
            log('    ' + d)

    # ---- cache ----------------------------------------------------------
    log('\n== 3/7 · inject_cache (al final del todo) ==')
    cache = inject_cache(carpeta, nombres)

    log('\n== 4/7 · verificación data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log(f"  con valor: {ver['con_valor']} · \"\" por diseño: "
        f"{ver['vacias_por_diseno']} · fallos: {len(ver['fallos'])}")
    for f in ver['fallos'][:10]:
        log('    ' + f)

    log('\n== 5/7 · sensibilidad pycel (2 inputs por fichero) ==')
    sens, sens_fallos = sensibilidad(carpeta, nombres)
    for s in sens:
        for p in s['pruebas']:
            log(f"  {s['fichero']} {p['input']}: {p['coste_antes']} → "
                f"{p['coste_despues']} €  {'OK' if p['cambia'] else 'NO CAMBIA'}")

    log('\n== 6/7 · censo-entregables --fail ==')
    cen = censo(carpeta)

    log('\n== 7/7 · demostraciones ==')
    demos = {
        'factor_17_filas': demo_factor(carpeta, respaldo or ORIGEN),
        'iva_en_celda': demo_iva(carpeta, nombres),
        'merma_por_categoria_fila_vacia': demo_merma(carpeta),
        'pases_02': demo_02(carpeta) if _hay_hoja(
            carpeta, '02-menu-degustacion.xlsx', '6. Pase') else None,
        'rotacion_03': demo_03(carpeta) if _hay_hoja(
            carpeta, '03-menu-del-dia.xlsx', 'Rotación Semanal') else None,
        'carrot_cake_07': demo_07(carpeta) if _hay_hoja(
            carpeta, '07-cafeteria-brunch.xlsx', 'Carrot Cake') else None,
        'tostada_aguacate_07': demo_tostada(carpeta) if '07-cafeteria-brunch.xlsx' in nombres else None,
        'proteccion': demo_proteccion(carpeta, nombres),
    }

    fallos = []
    for g in grupos:
        if hasattr(g, 'demos'):
            propias = g.demos(carpeta, ORIGEN, DEMOS)
            fallos += propias.pop('fallos', [])
            demos.update(propias)
            log(f'  demostraciones de {g.__name__}: {len(propias)} bloques')
    if idem.get('diferencias'):
        fallos.append(f"idempotencia: {idem['diferencias']} diferencias")
    if ver['fallos']:
        fallos.append(f"cache: {len(ver['fallos'])} fórmulas sin valor")
    fallos += sens_fallos
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió ' + str(cen['exit']))
    fallos += bono['fallos']

    informe = {
        'producto': 'kit-escandallos',
        'version': '2.0',
        'spec': SPEC,
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'carpeta_origen': ORIGEN,
        'carpeta_trabajo': carpeta,
        'grupos': pedidos,
        'ficheros': informe_ficheros,
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'data_only_formulas_nuevas': ver,
            'sensibilidad_pycel': sens,
            'censo_entregables': cen,
            'bono_guia': bono,
        },
        'demostraciones': demos,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1)
        log(f'\ninforme → {args.json}')

    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'TODO VERDE (idempotencia, cache, pycel, censo)'))
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (BÓRRALO sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
