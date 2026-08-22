#!/usr/bin/env python3
"""
grupo_a.py — §2 de `kit-escandallos-v2-SPEC.md`: lo específico de 01, 02, 03 y 07.

  · 02 menú degustación (DOM-16/TEC-17/COM-14/TEC-09): pestañas «6. Pase»…
    «9. Pase» vacías con la rejilla completa, Resumen ampliado a 9 referencias
    que ignoran los pases sin datos, y UN ÚNICO food cost objetivo editable en
    el Resumen del que cuelgan los nueve pases.
  · 03 menú del día (TEC-16/COM-15/DOM-17): hoja «Rotación Semanal» L-V con el
    coste, el PVP y el food cost de cada día, más la media de la semana. El
    food cost objetivo también se unifica (TEC-09: el Resumen decía 33 % y las
    tres pestañas 30 %).
  · 07 cafetería/brunch (DOM-08/TEC-23/COM-12): pestaña «Carrot Cake», por
    raciones (12 porciones por tanda).
  · 01: nada propio — sus dos correcciones (microgreens y el fondo oscuro) son
    del motor §1.7.

Se ejecuta en dos tiempos alrededor del motor:
  pre(wb, fname)  → crea las hojas nuevas en el layout v1 para que el motor las
                    trate como a cualquier otra (columna Factor, filas libres…).
  post(wb, fname) → cablea lo que depende del layout FINAL (referencias entre
                    hojas, food cost único, rotación semanal).
"""
import copy

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

import motor

FICHEROS = ['01-escandallo-estandar.xlsx', '02-menu-degustacion.xlsx',
            '03-menu-del-dia.xlsx', '07-cafeteria-brunch.xlsx']

PASES_NUEVOS = ['6. Pase', '7. Pase', '8. Pase', '9. Pase']
PASES = ['1. Aperitivo', '2. Entrante', '3. Pescado', '4. Carne', '5. Postre'] \
    + PASES_NUEVOS

DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

# ORQ-7.3 (SPEC §7.3): la primera versión de esta receta daba 0,6157 €/ración
# y un PVP de 2,42 € con IVA — por debajo de lo que cuesta una porción real y
# fuera de mercado para una vitrina. El problema no era la fórmula sino los
# GRAMAJES y los PRECIOS: media tarta y precios de harina a granel. Aquí va la
# tanda de un layer cake de 20 cm en dos capas (12 porciones) con precios de
# compra HORECA realistas: sale ≈ 1,10-1,15 €/ración y, con el food cost de
# vitrina del 25 %, un PVP ≈ 4,5 € sin IVA (≈ 4,95 € con IVA).
# El huevo va con merma 0: se mide en PIEZAS y la merma del 11 % es la cáscara
# (COM-M15); el motor lo forzaría igualmente, se deja explícito.
CARROT_CAKE = [
    # ingrediente, categoría, ud compra, precio, cantidad, ud uso, merma
    ('Zanahoria', 'Verdura raíz', 'kg', 1.40, 0.45, 'kg', 0.20),
    ('Harina floja', 'Secos/granos', 'kg', 1.20, 0.30, 'kg', 0.02),
    ('Azúcar moreno', 'Secos/granos', 'kg', 2.20, 0.28, 'kg', 0.02),
    ('Huevo campero', 'Huevos', 'docena', 4.20, 6, 'ud', 0),
    ('Aceite de girasol', 'Aceites/grasas', 'L', 2.40, 0.24, 'L', 0),
    ('Nueces peladas', 'Secos/granos', 'kg', 16.00, 0.12, 'kg', 0.05),
    ('Canela molida', 'Especias/hierbas', 'kg', 32.00, 0.008, 'kg', 0.02),
    ('Queso crema (frosting)', 'Lácteos', 'kg', 7.90, 0.45, 'kg', 0.03),
    ('Azúcar glas (frosting)', 'Secos/granos', 'kg', 2.80, 0.20, 'kg', 0.02),
    ('Mantequilla (frosting)', 'Lácteos', 'kg', 10.50, 0.12, 'kg', 0.03),
]
CARROT_RACIONES = 12
CARROT_FC = 0.25          # food cost objetivo de vitrina (SPEC §7.3)
CARROT_TITULO = ('Carrot Cake con Frosting de Queso — tanda de 12 raciones '
                 '(molde de 20 cm, dos capas)')


# ==========================================================================
# Utilidades
# ==========================================================================
def _mover_antes(wb, ws, nombre_destino):
    hojas = wb._sheets
    hojas.remove(ws)
    hojas.insert(hojas.index(wb[nombre_destino]), ws)


def _clonar(wb, origen, titulo, antes=None):
    """Clona una hoja de escandallo v1 y la deja VACÍA de ingredientes."""
    src = wb[origen]
    nueva = wb.copy_worksheet(src)
    nueva.title = titulo
    lay = motor.Layout(nueva)
    for fila in range(lay.d0, lay.d1 + 1):
        for col in range(1, 8):          # A..G del layout v1 (G = Merma)
            nueva.cell(row=fila, column=col).value = None
    if antes:
        _mover_antes(wb, nueva, antes)
    return nueva


def _fila_por(ws, prefijo, defecto, col=2):
    """Fila cuya celda de etiqueta empieza por `prefijo`; `defecto` si no está.

    Las plantillas cambian de altura entre v1 y v2 (el Resumen del 02 pasa de 5
    a 9 pases), así que capturar los estilos por número de fila fijo haría que
    la 2.ª pasada copiase el estilo de otra fila. Se buscan por etiqueta.
    """
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.startswith(prefijo):
            return r
    return defecto


def _verde(cel, fmt=None):
    cel.fill = PatternFill('solid', fgColor=motor.VERDE)
    cel.protection = Protection(locked=False)
    if fmt:
        cel.number_format = fmt


def _quitar_verde(cel):
    cel.fill = PatternFill()
    cel.protection = Protection(locked=True)


# ==========================================================================
# pre — hojas nuevas, todavía en layout v1
# ==========================================================================
def pre(wb, fname, informe):
    if fname == '02-menu-degustacion.xlsx':
        for i, titulo in enumerate(PASES_NUEVOS, start=6):
            if titulo in wb.sheetnames:
                continue
            ws = _clonar(wb, '1. Aperitivo', titulo, antes='Resumen')
            ws['A1'] = f'{titulo} — escribe aquí el nombre del pase'
            ws['A2'] = ('Pase libre: rellena los ingredientes y el Resumen lo '
                        'suma solo. Si lo dejas vacío, no cuenta.')
            informe.append(f'pestaña «{titulo}» creada (rejilla completa, vacía)')

    elif fname == '07-cafeteria-brunch.xlsx':
        if 'Carrot Cake' not in wb.sheetnames:
            ws = _clonar(wb, 'Tostada Aguacate', 'Carrot Cake')
            informe.append('pestaña «Carrot Cake» creada')
        ws = wb['Carrot Cake']
        ws['A1'] = CARROT_TITULO
        ws['A2'] = ('Escandallo por TANDA. El «Nº de raciones» reparte el coste '
                    'entre las porciones que salen.')
        lay = motor.Layout(ws)
        col_merma = 8 if lay.v2 else 7      # H en v2, G en v1
        for i, (nom, cat, uc, pre_, cant, uu, merma) in enumerate(CARROT_CAKE):
            fila = lay.d0 + i
            ws.cell(row=fila, column=1, value=nom)
            ws.cell(row=fila, column=2, value=cat)
            ws.cell(row=fila, column=3, value=uc)
            ws.cell(row=fila, column=4, value=pre_)
            ws.cell(row=fila, column=5, value=cant)
            ws.cell(row=fila, column=6, value=uu)
            ws.cell(row=fila, column=col_merma, value=merma)
        for fila in range(lay.d0 + len(CARROT_CAKE), lay.d1 + 1):
            for col in range(1, 7):
                ws.cell(row=fila, column=col).value = None
            ws.cell(row=fila, column=col_merma).value = None


# ==========================================================================
# post — cableado que depende del layout final
# ==========================================================================
def post(wb, fname, informe, registro):
    if fname == '02-menu-degustacion.xlsx':
        _resumen_02(wb, informe, registro)
    elif fname == '03-menu-del-dia.xlsx':
        _resumen_03(wb, informe, registro)
        _rotacion_03(wb, informe, registro)
    elif fname == '07-cafeteria-brunch.xlsx':
        ws = wb['Carrot Cake']
        b = motor.bloque(ws)
        ws.cell(row=b['raciones'], column=9, value=CARROT_RACIONES)
        registro.append((ws.title, f"I{b['raciones']}", CARROT_RACIONES))
        # ORQ-7.3: la vitrina de cafetería trabaja al 25 %, no al 30 % genérico
        # que hereda la hoja clonada de la tostada.
        cfc = ws.cell(row=b['fc_objetivo'], column=9)
        cfc.value = CARROT_FC
        cfc.number_format = motor.FMT_PCT
        informe.append(f'Carrot Cake: {CARROT_RACIONES} raciones por tanda y '
                       f'food cost de vitrina {CARROT_FC:.0%} (COSTE POR '
                       'RACIÓN activo; SPEC §7.3)')


def _fc_desde(wb, hojas, ref, informe):
    """Ata el «Food Cost objetivo» de cada hoja a una única celda editable."""
    for titulo in hojas:
        ws = wb[titulo]
        b = motor.bloque(ws)
        cel = ws.cell(row=b['fc_objetivo'], column=9)
        cel.value = f'={ref}'
        cel.number_format = motor.FMT_PCT
        _quitar_verde(cel)
        etiqueta = ws.cell(row=b['fc_objetivo'], column=1)
        etiqueta.value = 'Food Cost objetivo (%) — se edita en el Resumen'
    informe.append(f'food cost objetivo único: {len(hojas)} hojas leen {ref}')


# --------------------------------------------------------------------------
# 02 · Resumen de 9 pases
# --------------------------------------------------------------------------
def _resumen_02(wb, informe, registro):
    ws = wb['Resumen']
    est_lbl = copy.copy(ws.cell(row=5, column=2)._style)
    est_val = copy.copy(ws.cell(row=5, column=3)._style)
    f_tot = _fila_por(ws, 'COSTE TOTAL MENÚ', 10)
    f_fc = _fila_por(ws, 'Food Cost objetivo', 12)
    f_pvp = _fila_por(ws, 'PVP SUGERIDO', 13)
    est_tot_l = copy.copy(ws.cell(row=f_tot, column=2)._style)
    est_tot_v = copy.copy(ws.cell(row=f_tot, column=3)._style)
    est_in_l = copy.copy(ws.cell(row=f_fc, column=2)._style)
    est_in_v = copy.copy(ws.cell(row=f_fc, column=3)._style)
    est_pvp_l = copy.copy(ws.cell(row=f_pvp, column=2)._style)
    est_pvp_v = copy.copy(ws.cell(row=f_pvp, column=3)._style)
    fc_previo = ws.cell(row=f_fc, column=3).value
    if not isinstance(fc_previo, (int, float)):
        fc_previo = 0.25

    for r in range(5, ws.max_row + 1):
        for c in range(2, 4):
            ws.cell(row=r, column=c).value = None

    fila = 5
    for titulo in PASES:
        b = motor.bloque(wb[titulo])
        cl = ws.cell(row=fila, column=2, value=titulo)
        cl._style = copy.copy(est_lbl)
        cv = ws.cell(row=fila, column=3)
        # DOM-02: la celda correcta es COSTE POR RACIÓN, no COSTE TOTAL DEL
        # PLATO. Con «Nº de raciones» > 1 (una tanda de 10 aperitivos, un litro
        # de sopa para 8) el resumen multiplicaba el menú por las raciones.
        cv.value = (f"=IF('{titulo}'!$J${b['ingredientes']}=0,\"\","
                    f"'{titulo}'!$J${b['coste_racion']})")
        cv._style = copy.copy(est_val)
        cv.number_format = motor.FMT_EUR
        registro.append(('Resumen', f'C{fila}', cv.value))
        fila += 1

    r_tot = fila
    ws.cell(row=r_tot, column=2, value='COSTE TOTAL MENÚ')._style = copy.copy(est_tot_l)
    c = ws.cell(row=r_tot, column=3, value=f'=SUM(C5:C{r_tot - 1})')
    c._style = copy.copy(est_tot_v)
    c.number_format = motor.FMT_EUR
    registro.append(('Resumen', f'C{r_tot}', c.value))

    r_fc = r_tot + 2
    ws.cell(row=r_fc, column=2,
            value='Food Cost objetivo (%) — el único food cost del libro: '
                  'los pases lo leen de aquí')._style = copy.copy(est_in_l)
    c = ws.cell(row=r_fc, column=3, value=fc_previo)
    c._style = copy.copy(est_in_v)
    c.number_format = motor.FMT_PCT
    _verde(c, motor.FMT_PCT)

    r_pvp = r_fc + 1
    ws.cell(row=r_pvp, column=2, value='PVP SUGERIDO (sin IVA)')._style = copy.copy(est_pvp_l)
    c = ws.cell(row=r_pvp, column=3,
                value=f'=IFERROR(C{r_tot}/C{r_fc},"")')
    c._style = copy.copy(est_pvp_v)
    c.number_format = motor.FMT_EUR
    registro.append(('Resumen', f'C{r_pvp}', c.value))

    r_iva = r_pvp + 1
    ws.cell(row=r_iva, column=2, value=motor.ETIQ_IVA)._style = copy.copy(est_in_l)
    c = ws.cell(row=r_iva, column=3, value=0.10)
    c._style = copy.copy(est_in_v)
    c.number_format = motor.FMT_PCT
    _verde(c, motor.FMT_PCT)

    r_civ = r_iva + 1
    ws.cell(row=r_civ, column=2, value='PVP CON IVA')._style = copy.copy(est_pvp_l)
    c = ws.cell(row=r_civ, column=3,
                value=f'=IFERROR(C{r_pvp}*(1+C{r_iva}),"")')
    c._style = copy.copy(est_pvp_v)
    c.number_format = motor.FMT_EUR
    registro.append(('Resumen', f'C{r_civ}', c.value))

    nota = ws.cell(row=r_civ + 2, column=2,
                   value='Los pases sin ingredientes no suman: el menú puede '
                         'tener de 5 a 9 pases sin tocar una sola fórmula.')
    nota.font = Font(italic=True, size=9)

    _fc_desde(wb, PASES, f'Resumen!$C${r_fc}', informe)
    informe.append(f'Resumen ampliado a 9 pases (C5:C{r_tot - 1}) con '
                   f'IF(pase vacío,"",…)')


# --------------------------------------------------------------------------
# 03 · Resumen Menú + Rotación Semanal
# --------------------------------------------------------------------------
PLATOS_03 = ['Primer Plato', 'Segundo Plato', 'Postre']


def _resumen_03(wb, informe, registro):
    ws = wb['Resumen Menú']
    est_lbl = copy.copy(ws.cell(row=5, column=2)._style)
    est_val = copy.copy(ws.cell(row=5, column=3)._style)
    f_tot = _fila_por(ws, 'COSTE TOTAL MENÚ', 11)
    f_fc = _fila_por(ws, 'Food Cost objetivo', 13)
    f_pvp = _fila_por(ws, 'PVP SUGERIDO', 14)
    est_tot_l = copy.copy(ws.cell(row=f_tot, column=2)._style)
    est_tot_v = copy.copy(ws.cell(row=f_tot, column=3)._style)
    est_in_l = copy.copy(ws.cell(row=f_fc, column=2)._style)
    est_in_v = copy.copy(ws.cell(row=f_fc, column=3)._style)
    est_pvp_l = copy.copy(ws.cell(row=f_pvp, column=2)._style)
    est_pvp_v = copy.copy(ws.cell(row=f_pvp, column=3)._style)
    fc_previo = ws.cell(row=f_fc, column=3).value
    if not isinstance(fc_previo, (int, float)):
        fc_previo = 0.33
    extras = []
    for r in (8, 9, 10):
        v = ws.cell(row=r, column=3).value
        extras.append(v if isinstance(v, (int, float)) else 0)

    for r in range(5, ws.max_row + 1):
        for c in range(2, 4):
            ws.cell(row=r, column=c).value = None

    fila = 5
    for titulo in PLATOS_03:
        b = motor.bloque(wb[titulo])
        ws.cell(row=fila, column=2, value=titulo)._style = copy.copy(est_lbl)
        # DOM-02: COSTE POR RACIÓN, no COSTE TOTAL DEL PLATO.
        c = ws.cell(row=fila, column=3,
                    value=f"='{titulo}'!$J${b['coste_racion']}")
        c._style = copy.copy(est_val)
        c.number_format = motor.FMT_EUR
        registro.append(('Resumen Menú', f'C{fila}', c.value))
        fila += 1
    for etiqueta, valor in zip(('Pan (por comensal)', 'Bebida (agua/vino copa)',
                                'Café'), extras):
        ws.cell(row=fila, column=2, value=etiqueta)._style = copy.copy(est_lbl)
        c = ws.cell(row=fila, column=3, value=valor)
        c._style = copy.copy(est_val)
        c.number_format = motor.FMT_EUR
        _verde(c, motor.FMT_EUR)
        fila += 1

    r_tot = fila
    ws.cell(row=r_tot, column=2,
            value='COSTE TOTAL MENÚ DEL DÍA')._style = copy.copy(est_tot_l)
    c = ws.cell(row=r_tot, column=3, value=f'=SUM(C5:C{r_tot - 1})')
    c._style = copy.copy(est_tot_v)
    c.number_format = motor.FMT_EUR
    registro.append(('Resumen Menú', f'C{r_tot}', c.value))

    r_fc = r_tot + 2
    ws.cell(row=r_fc, column=2,
            value='Food Cost objetivo (%) — el único food cost del libro: '
                  'los platos lo leen de aquí')._style = copy.copy(est_in_l)
    c = ws.cell(row=r_fc, column=3, value=fc_previo)
    c._style = copy.copy(est_in_v)
    c.number_format = motor.FMT_PCT
    _verde(c, motor.FMT_PCT)

    r_pvp = r_fc + 1
    ws.cell(row=r_pvp, column=2,
            value='PVP SUGERIDO MENÚ (sin IVA)')._style = copy.copy(est_pvp_l)
    c = ws.cell(row=r_pvp, column=3, value=f'=IFERROR(C{r_tot}/C{r_fc},"")')
    c._style = copy.copy(est_pvp_v)
    c.number_format = motor.FMT_EUR
    registro.append(('Resumen Menú', f'C{r_pvp}', c.value))

    r_iva = r_pvp + 1
    ws.cell(row=r_iva, column=2, value=motor.ETIQ_IVA)._style = copy.copy(est_in_l)
    c = ws.cell(row=r_iva, column=3, value=0.10)
    c._style = copy.copy(est_in_v)
    c.number_format = motor.FMT_PCT
    _verde(c, motor.FMT_PCT)

    r_civ = r_iva + 1
    ws.cell(row=r_civ, column=2, value='PVP CON IVA')._style = copy.copy(est_pvp_l)
    c = ws.cell(row=r_civ, column=3, value=f'=IFERROR(C{r_pvp}*(1+C{r_iva}),"")')
    c._style = copy.copy(est_pvp_v)
    c.number_format = motor.FMT_EUR
    registro.append(('Resumen Menú', f'C{r_civ}', c.value))

    _fc_desde(wb, PLATOS_03, f"'Resumen Menú'!$C${r_fc}", informe)
    informe.append('Resumen Menú con IVA en celda y food cost único')
    return {'fc': r_fc, 'pvp': r_pvp, 'extras': (8, 9, 10), 'total': r_tot}


ROT_HDR = ['Día', 'Primer plato', 'Coste (€)', 'Segundo plato', 'Coste (€)',
           'Postre', 'Coste (€)', 'Extras (pan/bebida/café)',
           'COSTE MENÚ (€)', 'PVP menú (del Resumen)', 'FOOD COST (%)']
ROT_ANCHOS = [12, 26, 12, 26, 12, 24, 12, 22, 16, 18, 14]


def _rotacion_03(wb, informe, registro):
    ws_res = wb['Resumen Menú']
    r_fc = r_pvp = None
    for r in range(1, ws_res.max_row + 1):
        v = ws_res.cell(row=r, column=2).value
        if isinstance(v, str) and v.startswith('Food Cost objetivo'):
            r_fc = r
        elif isinstance(v, str) and v.startswith('PVP SUGERIDO MENÚ'):
            r_pvp = r
    r_ex0 = 8

    nueva = 'Rotación Semanal' not in wb.sheetnames
    ws = wb['Rotación Semanal'] if not nueva else wb.create_sheet('Rotación Semanal')
    if nueva:
        wb._sheets.remove(ws)
        wb._sheets.insert(wb._sheets.index(ws_res) + 1, ws)
    for r in range(1, ws.max_row + 1):
        for c in range(1, 12):
            ws.cell(row=r, column=c).value = None
    # sin esto la 2.ª pasada apila otra regla idéntica sobre K5:K11
    ws.conditional_formatting = ConditionalFormattingList()

    ws['A1'] = 'Rotación Semanal — los 5 menús del día'
    ws['A1'].font = Font(bold=True, size=13)
    # COM-B19: el color tenía que decir lo mismo que el texto. El lunes está
    # ENLAZADO a las pestañas de escandallo (gris, bloqueado) y la columna
    # «PVP menú» es un RESULTADO que viene del Resumen: pintarla de verde la
    # anunciaba como editable en una hoja protegida donde no se puede escribir.
    ws['A2'] = ('Código de color: VERDE = lo escribes tú. GRIS = viene de otra '
                'hoja, no se toca. El LUNES trae el nombre y el coste '
                'enlazados a las pestañas «Primer Plato», «Segundo Plato» y '
                '«Postre»: cambia la receta allí y aquí cambia solo (por eso '
                'sus tres celdas de coste están en gris). De MARTES a VIERNES '
                'escribes tú el nombre y el coste de cada plato en las celdas '
                'verdes. La columna «PVP menú» también es gris: es el precio '
                'único del menú y sale del «Resumen Menú», así que el food '
                'cost es lo que cambia cada día. ¿Cobras otro precio un día '
                'concreto (festivo)? Revisar → Desproteger hoja y escribe '
                'encima de esa celda.')
    ws['A2'].font = Font(size=9, italic=True)

    negrita = Font(bold=True, color='FFFFFF')
    fondo = PatternFill('solid', fgColor=motor.CAB)
    for i, t in enumerate(ROT_HDR, start=1):
        c = ws.cell(row=4, column=i, value=t)
        c.font = negrita
        c.fill = fondo
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ROT_ANCHOS[i - 1]

    nombres_lunes = []
    for titulo in PLATOS_03:
        a1 = wb[titulo].cell(row=1, column=1).value or titulo
        nombres_lunes.append(a1.split('—')[-1].strip() if '—' in a1 else a1)

    extras_ref = (f"='Resumen Menú'!$C${r_ex0}+'Resumen Menú'!$C${r_ex0 + 1}"
                  f"+'Resumen Menú'!$C${r_ex0 + 2}")
    for i, dia in enumerate(DIAS):
        fila = 5 + i
        ws.cell(row=fila, column=1, value=dia).font = Font(bold=True)
        for j, titulo in enumerate(PLATOS_03):
            col_nom = 2 + j * 2
            col_cos = 3 + j * 2
            cn = ws.cell(row=fila, column=col_nom)
            cc = ws.cell(row=fila, column=col_cos)
            if i == 0:
                cn.value = nombres_lunes[j]
                _verde(cn)     # el nombre sí se cambia; el coste es fórmula
                b = motor.bloque(wb[titulo])
                # DOM-02: COSTE POR RACIÓN, no COSTE TOTAL DEL PLATO.
                cc.value = f"='{titulo}'!$J${b['coste_racion']}"
                registro.append(('Rotación Semanal',
                                 f'{get_column_letter(col_cos)}{fila}', cc.value))
            else:
                _verde(cn)
                _verde(cc, motor.FMT_EUR)
            cc.number_format = motor.FMT_EUR
        ce = ws.cell(row=fila, column=8, value=extras_ref)
        ce.number_format = motor.FMT_EUR
        registro.append(('Rotación Semanal', f'H{fila}', ce.value))
        ci = ws.cell(row=fila, column=9,
                     value=f'=IF(SUM(C{fila},E{fila},G{fila})=0,"",'
                           f'SUM(C{fila},E{fila},G{fila},H{fila}))')
        ci.number_format = motor.FMT_EUR
        ci.font = Font(bold=True)
        registro.append(('Rotación Semanal', f'I{fila}', ci.value))
        cp = ws.cell(row=fila, column=10, value=f"='Resumen Menú'!$C${r_pvp}")
        cp.number_format = motor.FMT_EUR
        _quitar_verde(cp)          # COM-B19: es un resultado, no una entrada
        registro.append(('Rotación Semanal', f'J{fila}', cp.value))
        ck = ws.cell(row=fila, column=11,
                     value=f'=IFERROR(I{fila}/J{fila},"")')
        ck.number_format = motor.FMT_PCT1
        registro.append(('Rotación Semanal', f'K{fila}', ck.value))

    r_med = 11
    ws.cell(row=r_med, column=1, value='MEDIA SEMANA').font = Font(bold=True)
    for col, formula, fmt in (
            (9, '=IFERROR(SUM(I5:I9)/COUNTIF(I5:I9,">0"),"")', motor.FMT_EUR),
            (10, '=IFERROR(SUM(J5:J9)/COUNTIF(J5:J9,">0"),"")', motor.FMT_EUR),
            (11, f'=IFERROR(I{r_med}/J{r_med},"")', motor.FMT_PCT1)):
        c = ws.cell(row=r_med, column=col, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor=motor.ORO)
        registro.append(('Rotación Semanal',
                         f'{get_column_letter(col)}{r_med}', formula))

    ws.conditional_formatting.add(
        f'K5:K{r_med}',
        FormulaRule(formula=[f"AND(ISNUMBER(K5),K5>'Resumen Menú'!$C${r_fc})"],
                    fill=PatternFill('solid', start_color=motor.ROJO_BG,
                                     end_color=motor.ROJO_BG),
                    font=Font(color=motor.ROJO_FG, bold=True)))
    informe.append('hoja «Rotación Semanal» (L-V, coste/PVP/food cost por día '
                   '+ media semanal)')
