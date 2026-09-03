#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_plan-accion-90-dias.py — Libro 8 del pack «Guía Food Cost + Ingeniería de
Menú» (SPEC §2.2, fila 8).

Genera `build/plan-accion-90-dias.xlsx`:

  Instrucciones · Decisiones · Calendario 90 Días · KPI de Seguimiento

Es la hoja que le pone FECHA y RESPONSABLE a lo que sale de las otras siete
herramientas del pack. No mide disciplina diaria: mide decisiones de carta a lo
largo de un trimestre.

⚠️ NO es el plan de 4 semanas del bono del Kit de Escandallos. Aquel es diario y
va de crear el hábito de escandallar; éste es trimestral y va de decidir qué
pasa con cada plato. La hoja «Instrucciones» lo declara en voz alta para que
nadie compre dos veces lo mismo ni intente cruzarlos.

DECISIONES TÉCNICAS
-------------------
* Nada de `SUMIF`/`COUNTIF` con una celda de criterio que pueda venir vacía:
  pycel evalúa los dos brazos de un `IF` de forma ansiosa y el guardián no
  protege — revienta con «Couldn't parse criteria: None» y deja el libro entero
  sin cache. Aquí todos los criterios son etiquetas FIJAS (los desplegables
  cerrados de herramienta, decisión, estado y bloque), así que `SUMIF` y
  `COUNTIF` son seguros; los cruces de dos condiciones van con `SUMPRODUCT`.
* Funciones prohibidas (INDIRECT, COUNTA, PMT, OFFSET, LET, LAMBDA, XLOOKUP,
  matrices dinámicas): cero. Tampoco `COUNT`; su papel lo hace
  `COUNTIF(rango,"<>")`.
* Objetivo de cierre y sentido de cada KPI («¿bajar es bueno?») en celdas
  verdes: ninguna constante dentro de una fórmula.
* «Sin dato» = `""`, nunca `0`; `IFERROR(...,"")` en todo cociente.

Salida fija (sin argumentos): `<carpeta>/build/plan-accion-90-dias.xlsx`
"""
import json
import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

motor.CTX['producto'] = 'guia-food-cost-ingenieria-menu'

# --------------------------------------------------------------------------
NOMBRE = 'plan-accion-90-dias'
TITULO_LIBRO = 'Plan de Acción a 90 Días'
SUBTITULO = 'AI Chef Pro · aichef.pro — Guía Food Cost + Ingeniería de Menú'
SUBJECT = 'Guía Food Cost + Ingeniería de Menú · Versión 1.0 · septiembre 2026'
VERSION = ('Versión 1.0 · septiembre 2026 · '
           'aichef.pro/guia-food-cost-ingenieria-menu · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 · '
       'johnguerrero.es')
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde, '
               'desprotege la hoja (sin contraseña).')
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

N = motor.NARROW
GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL = 'FFF8E1', '1565C0'
EUR, PCT, ENT = motor.FMT_EUR, motor.FMT_PCT, motor.FMT_ENT
FECHA = motor.FMT_FECHA

HERRAMIENTAS = ['Ficha de escandallo', 'Test de rendimiento',
                'Precio objetivo', 'Matriz multi-método',
                'Simulador multicanal', 'Carta de bebidas',
                'Cuadro de mando prime cost']
# Mismo vocabulario que emite `matriz-multimetodo-carta!Comparativa!J`
# (columna «Decisión sugerida»): «Revisar» es una de sus seis salidas y tiene
# que estar en esta lista o el plan no puede citar la decisión tal cual.
DECISIONES = ['Reformular', 'Resubir', 'Rediseñar', 'Retirar', 'Revisar',
              'Negociar', 'Mantener']
ESTADOS = ['Pendiente', 'En curso', 'Hecha', 'Descartada']
SI_NO = ['Sí', 'No']

DF0, DF1 = 5, 24          # Decisiones: 10 de ejemplo + 10 filas libres
CF0, CF1 = 5, 30          # Calendario: 13 semanas × 2 hitos
KF0, KF1 = 5, 12          # KPI: 5 sembrados + 3 filas libres

BLOQUES = ['Medir', 'Escandallar y clasificar', 'Aplicar decisiones',
           'Medir el efecto', 'Revisión']
RESPONSABLES = ['Gerente', 'Jefe de cocina', 'Jefe de sala', 'Jefe de compras']

# 13 semanas × 2 hitos. (semana, bloque, hito, responsable)
CALENDARIO = [
    (1, 'Medir', 'Sacar el informe de ventas por plato del último mes y '
                 'volcarlo en la hoja «Datos» de la matriz.', 'Gerente'),
    (1, 'Medir', 'Hacer inventario de cierre y anotar stock inicial, compras y '
                 'stock final del mes.', 'Jefe de cocina'),
    (2, 'Medir', 'Test de rendimiento de los 10 productos que más pesan en la '
                 'compra, con báscula y hoja de rendimiento.', 'Jefe de cocina'),
    (2, 'Medir', 'Medir la merma de cocción de las cinco elaboraciones que más '
                 'se venden.', 'Jefe de cocina'),
    (3, 'Escandallar y clasificar',
        'Escandallar los platos de más venta con las mermas MEDIDAS, no con '
        'las de referencia.', 'Jefe de cocina'),
    (3, 'Escandallar y clasificar',
        'Rellenar coste por ración y PVP sin IVA de toda la carta en la '
        'matriz multi-método.', 'Gerente'),
    (4, 'Escandallar y clasificar',
        'Leer las cuatro clasificaciones y la hoja «Comparativa» plato a '
        'plato.', 'Gerente'),
    (4, 'Escandallar y clasificar',
        'Pasar a la hoja «Decisiones» los platos con tres o cuatro lecturas '
        'fuera de la mejor categoría.', 'Gerente'),
    (5, 'Aplicar decisiones',
        'Aplicar las subidas de precio decididas y reimprimir la carta.',
        'Gerente'),
    (5, 'Aplicar decisiones',
        'Explicar a sala los cambios de precio y qué platos hay que sugerir.',
        'Jefe de sala'),
    (6, 'Aplicar decisiones',
        'Reformular las recetas marcadas: ficha de escandallo nueva y prueba '
        'en cocina.', 'Jefe de cocina'),
    (6, 'Aplicar decisiones',
        'Negociar con los proveedores de las tres partidas que más pesan en la '
        'compra.', 'Jefe de compras'),
    (7, 'Aplicar decisiones',
        'Rediseñar la carta: orden, agrupación y sitio de los platos que hay '
        'que empujar.', 'Gerente'),
    (7, 'Aplicar decisiones',
        'Excluir del delivery los platos que no son viables al precio techo de '
        'la aplicación.', 'Gerente'),
    (8, 'Aplicar decisiones',
        'Retirar los platos condenados y ajustar la compra de la semana '
        'siguiente.', 'Jefe de cocina'),
    (8, 'Aplicar decisiones',
        'Revisar la carta de bebidas: copa, botella y barril con su beverage '
        'cost.', 'Jefe de sala'),
    (9, 'Medir el efecto',
        'Cerrar el mes 1: ventas, compras e inventario; food cost real y prime '
        'cost.', 'Gerente'),
    (9, 'Medir el efecto',
        'Comparar el mix de ventas con el del mes 0: ¿se ha movido lo que '
        'queríamos mover?', 'Gerente'),
    (10, 'Medir el efecto',
         'Recalcular la matriz con las ventas nuevas y ver qué platos han '
         'cambiado de casilla.', 'Gerente'),
    (10, 'Medir el efecto',
         'Anotar en «Decisiones» el estado real de cada una y el impacto '
         'conseguido.', 'Gerente'),
    (11, 'Medir el efecto',
         'Re-escandallar las partidas cuyo proveedor haya subido precio en el '
         'trimestre.', 'Jefe de cocina'),
    (11, 'Medir el efecto',
         'Revisar packaging y comisión de las plataformas de delivery contra '
         'el contrato firmado.', 'Jefe de compras'),
    (12, 'Medir el efecto',
         'Cerrar el mes 2 y actualizar la hoja de KPI de seguimiento.',
         'Gerente'),
    (12, 'Medir el efecto',
         'Detectar las decisiones bloqueadas y darles fecha nueva o '
         'descartarlas con motivo.', 'Gerente'),
    (13, 'Revisión',
         'Cerrar el mes 3, actualizar los KPI y compararlos con el mes 0.',
         'Gerente'),
    (13, 'Revisión',
         'Decidir el trimestre siguiente: qué se mantiene, qué se repite y qué '
         'entra nuevo.', 'Gerente'),
]

# KPI: (nombre, formato, ¿bajar es bueno?, nota)
KPI_META = {
    'Food cost (%)': (PCT, 'Sí', 'Coste de producto sobre ventas netas de '
                                 'comida. Objetivo de la casa en el cuadro de '
                                 'mando.'),
    'Prime cost (%)': (PCT, 'Sí', 'Producto + personal con Seguridad Social, '
                                  'sobre ventas netas.'),
    'Ticket medio sin IVA (€)': (EUR, 'No', 'Ventas netas ÷ cubiertos. Sube si '
                                            'la carta empuja bien.'),
    'Margen de contribución por cubierto (€)': (
        EUR, 'No', 'El euro que deja cada comensal para pagar personal, '
                   'alquiler y luz.'),
    'Platos en carta (n.º)': (ENT, 'Sí', 'La poda es un objetivo, no un '
                                         'accidente. Cámbialo a «No» si tu '
                                         'plan es ampliar carta.'),
}


# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, landscape=True):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8


def encabezados(ws, fila, cols, alto=40):
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def bloque(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)


def total(ws, coord, contenido, fmt=None, formula=False):
    cel = (motor.f(ws, coord, contenido, fmt=fmt, bold=True) if formula
           else motor.val(ws, coord, contenido, fmt=fmt, bold=True))
    cel.fill = PatternFill('solid', fgColor=CREMA)
    return cel


def verde_propio(ws, fila, col_et, col_val, etiqueta, valor, fmt, nota,
                 col_nota=None):
    motor.val(ws, col_et + str(fila), etiqueta)
    motor.val(ws, col_val + str(fila), valor, fmt=fmt, verde_=True)
    if col_nota:
        motor.val(ws, col_nota + str(fila), nota)
    return '$' + col_val + '$' + str(fila)


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. Abre «Decisiones» y vuelca ahí lo que han dicho las otras herramientas '
    'del pack: qué plato, de qué hoja sale y qué se decide hacer con él.',
    '2. Ponle a cada decisión un responsable con nombre y apellido y una '
    'semana del plan (1-13). La «Fecha objetivo» la calcula el libro solo '
    'desde la «Fecha de inicio del plan» de la celda verde D36: una decisión '
    'sin dueño y sin fecha no se ejecuta.',
    '3. Estima el impacto en euros al mes. No hace falta precisión: hace falta '
    'orden de magnitud para saber por dónde empezar.',
    '4. Cambia el «Estado» según avances. El libro cuenta cuántas has cerrado, '
    'cuánto impacto llevas conseguido y cuánto queda pendiente.',
    '5. En «Calendario 90 Días» tienes 13 semanas con dos hitos cada una. Son '
    'una propuesta: reescríbelos con lo que toque en tu casa.',
    '6. Marca «Hecho» = Sí a medida que cierras hitos; el avance del plan y el '
    'de cada bloque se calculan solos.',
    '7. En «KPI de Seguimiento» anota la foto del mes 0 y luego los meses 1, 2 '
    'y 3. La lectura te dice si cada KPI mejora o empeora.',
    '8. A los 90 días, vuelve a la matriz con las ventas nuevas: los platos '
    'habrán cambiado de casilla y el ciclo empieza otra vez.',
]

NOTAS = [
    'ESTE PLAN NO ES EL PLAN DE 4 SEMANAS DEL BONO DEL KIT DE ESCANDALLOS. '
    'Aquel es DIARIO y sirve para coger el hábito de escandallar y de medir. '
    'Éste es TRIMESTRAL y sirve para decidir qué pasa con cada plato de la '
    'carta. Se complementan, pero no se sustituyen ni se solapan: si tienes los '
    'dos, primero el de 4 semanas para tener datos, después éste para decidir '
    'con ellos.',
    'Las 10 decisiones sembradas son un EJEMPLO de cómo se rellena, salidas de '
    'la carta modelada del pack. Bórralas y pon las tuyas.',
    'La columna «Decisión sugerida» de la hoja «Comparativa» de la matriz es '
    'una sugerencia automática, no una orden. Aquí decide una persona: puedes '
    'retirar un plato que la matriz pedía conservar si te ocupa cámara, mano de '
    'obra o sitio en la carta. Lo que no vale es decidirlo sin mirar el dato.',
    'El impacto en euros al mes es una ESTIMACIÓN tuya, no un cálculo del '
    'libro: la suma sirve para priorizar, no para presupuestar.',
    'El calendario arranca con todos los hitos en «No». Es correcto: un plan '
    'recién abierto tiene el 0' + N + '% de avance.',
    'En los KPI expresados en porcentaje, la variación son PUNTOS '
    'porcentuales: la resta directa entre el mes y el mes 0, no un porcentaje '
    'sobre el mes 0.',
]

NOTA_IVA_LIBRO = (
    'Todos los importes de este libro van SIN IVA, igual que en el resto del '
    'pack: el food cost y el margen se miden sobre la base imponible. El IVA '
    'repercutido no es ingreso y el IVA soportado se deduce en el modelo 303, '
    'así que ni uno ni otro entran en estos KPI.')


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila), NOTA_IVA_LIBRO, wrap=True)
    fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, landscape=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Decisiones»
# --------------------------------------------------------------------------
def hoja_decisiones(wb):
    ws = wb.create_sheet('Decisiones')
    cabecera(ws, 'Decisiones — qué se hace, quién y cuándo')
    apunte(ws, 'E3', 'Cada fila sale de una hoja del pack. Sin responsable y '
                     'sin fecha, no es una decisión: es una opinión.')
    encabezados(ws, 4, [
        ('A', '#', 5), ('B', 'Plato o área', 40),
        ('C', 'Herramienta de origen', 24), ('D', 'Decisión', 15),
        ('E', 'Responsable', 18), ('F', 'Semana del plan', 11),
        ('G', 'Fecha objetivo', 14), ('H', 'Estado', 14),
        ('I', 'Impacto estimado (€/mes)', 14),
        ('J', 'Impacto ya conseguido (€/mes)', 14), ('K', 'Notas', 40),
    ])
    ws.freeze_panes = 'B5'

    v_her, v_dec, v_est, v_sem, v_eur = [], [], [], [], []
    for i in range(DF1 - DF0 + 1):
        r = DF0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        if i < len(DE.DECISIONES_EJEMPLO):
            area, herramienta, decision, resp, semana, impacto = \
                DE.DECISIONES_EJEMPLO[i]
            motor.val(ws, f'B{r}', area)
            motor.val(ws, f'C{r}', herramienta)
            motor.val(ws, f'D{r}', decision)
            motor.val(ws, f'E{r}', resp)
            motor.val(ws, f'F{r}', semana, fmt=ENT)
            motor.val(ws, f'H{r}', ESTADOS[0])
            motor.val(ws, f'I{r}', impacto, fmt=EUR)
        else:
            ws[f'F{r}'].number_format = ENT
            ws[f'I{r}'].number_format = EUR
        motor.verde(ws, f'B{r}:F{r}')
        motor.verde(ws, f'H{r}:I{r}')
        motor.verde(ws, f'K{r}')
        v_her.append(f'C{r}')
        v_dec.append(f'D{r}')
        v_est.append(f'H{r}')
        v_sem.append(f'F{r}')
        v_eur.append(f'I{r}')
        motor.f(ws, f'G{r}',
                f'=IFERROR(IF(OR($F{r}="",$D$36=""),"",'
                f'$D$36+7*($F{r}-1)),"")', fmt=FECHA)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF(OR($B{r}="",$I{r}=""),"",'
                f'IF($H{r}="Hecha",$I{r},0)),"")', fmt=EUR)
        ws[f'J{r}'].font = Font(bold=True, color=AZUL)

    motor.dv_lista(ws, v_her, HERRAMIENTAS, titulo='Herramienta no válida')
    motor.dv_lista(ws, v_dec, DECISIONES, titulo='Decisión no válida')
    motor.dv_lista(ws, v_est, ESTADOS, titulo='Estado no válido')
    motor.dv_numerica(ws, v_sem, minimo=1, maximo=13,
                      titulo='Semana del plan',
                      mensaje='El plan tiene 13 semanas: escribe un número '
                              'entre 1 y 13.')
    motor.dv_numerica(ws, v_eur, minimo=0, titulo='Impacto estimado',
                      mensaje='Escribe el impacto estimado en euros al mes '
                              '(0 o más).')
    motor.semaforo_texto(ws, f'H{DF0}:H{DF1}', (
        ('Hecha', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('En curso', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Descartada', motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('Pendiente', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    # --- Resumen ----------------------------------------------------------
    bloque(ws, 'A26', 'RESUMEN Y PARÁMETROS — lo calcula el libro')
    motor.val(ws, 'B27', 'Decisiones registradas')
    motor.f(ws, 'D27', f'=COUNTIF($B${DF0}:$B${DF1},"<>")', fmt=ENT)
    motor.val(ws, 'B28', 'Decisiones cerradas (estado «Hecha»)')
    motor.f(ws, 'D28', f'=COUNTIF($H${DF0}:$H${DF1},"Hecha")', fmt=ENT)
    motor.val(ws, 'B29', 'Decisiones descartadas')
    motor.f(ws, 'D29', f'=COUNTIF($H${DF0}:$H${DF1},"Descartada")', fmt=ENT)
    motor.val(ws, 'B30', 'Decisiones cerradas (%)', bold=True)
    motor.f(ws, 'D30', '=IFERROR(IF($D$27=0,"",$D$28/$D$27),"")', fmt=PCT,
            bold=True)
    total(ws, 'B31', 'IMPACTO TOTAL ESTIMADO (€/mes)')
    total(ws, 'D31', f'=IFERROR(SUM($I${DF0}:$I${DF1}),"")', fmt=EUR,
          formula=True)
    motor.val(ws, 'B32', 'Impacto ya conseguido (€/mes)')
    motor.f(ws, 'D32',
            f'=IFERROR(SUMIF($H${DF0}:$H${DF1},"Hecha",$I${DF0}:$I${DF1}),"")',
            fmt=EUR)
    motor.val(ws, 'B33', 'Impacto pendiente (€/mes)')
    motor.f(ws, 'D33',
            '=IFERROR(IF(OR($D$31="",$D$32=""),"",$D$31-$D$32),"")', fmt=EUR)
    motor.val(ws, 'B34', 'Impacto total estimado a 12 meses (€)')
    motor.f(ws, 'D34', '=IFERROR(IF($D$31="","",$D$31*12),"")', fmt=EUR)
    p_obj = verde_propio(
        ws, 35, 'B', 'D', 'Objetivo de cierre a 90 días (%)', 0.80, PCT,
        'Qué parte de las decisiones te comprometes a cerrar en el trimestre. '
        'Si el porcentaje cerrado se queda por debajo, la celda se pone en '
        'rojo.', col_nota='F')
    motor.dv_porcentaje(ws, ['D35'], titulo='Objetivo de cierre',
                        prompt='Se escribe en tanto por uno: 0,80 = 80' + N +
                               '%.')
    motor.regla_expresion(ws, 'D30', f'=AND(ISNUMBER($D$30),$D$30<{p_obj})')
    p_fecha_inicio = verde_propio(
        ws, 36, 'B', 'D', 'Fecha de inicio del plan', date(2026, 9, 1), FECHA,
        'La semana 1 empieza este día; de aquí salen las fechas objetivo de '
        'las decisiones y del calendario. Cámbiala por tu fecha real.',
        col_nota='F')
    motor.dv_fecha(ws, ['D36'])

    bloque(ws, 'A37', 'RESUMEN POR TIPO DE DECISIÓN — lo calcula el libro')
    encabezados(ws, 38, [
        ('B', 'Decisión', None), ('C', 'Decisiones', None),
        ('D', 'Impacto estimado (€/mes)', None),
        ('E', 'Impacto conseguido (€/mes)', None),
    ], alto=34)
    for k, dec in enumerate(DECISIONES):
        r = 39 + k
        motor.val(ws, f'B{r}', dec, bold=True)
        motor.f(ws, f'C{r}', f'=COUNTIF($D${DF0}:$D${DF1},$B{r})', fmt=ENT)
        motor.f(ws, f'D{r}',
                f'=IFERROR(SUMIF($D${DF0}:$D${DF1},$B{r},'
                f'$I${DF0}:$I${DF1}),"")', fmt=EUR)
        motor.f(ws, f'E{r}',
                f'=IFERROR(SUMPRODUCT(--($D${DF0}:$D${DF1}=$B{r}),'
                f'--($H${DF0}:$H${DF1}="Hecha"),$I${DF0}:$I${DF1}),"")',
                fmt=EUR)

    bloque(ws, 'A46', 'RESUMEN POR HERRAMIENTA DE ORIGEN — lo calcula el libro')
    encabezados(ws, 47, [
        ('B', 'Herramienta de origen', None), ('C', 'Decisiones', None),
        ('D', 'Impacto estimado (€/mes)', None),
        ('E', 'Impacto conseguido (€/mes)', None),
    ], alto=34)
    for k, her in enumerate(HERRAMIENTAS):
        r = 48 + k
        motor.val(ws, f'B{r}', her, bold=True)
        motor.f(ws, f'C{r}', f'=COUNTIF($C${DF0}:$C${DF1},$B{r})', fmt=ENT)
        motor.f(ws, f'D{r}',
                f'=IFERROR(SUMIF($C${DF0}:$C${DF1},$B{r},'
                f'$I${DF0}:$I${DF1}),"")', fmt=EUR)
        motor.f(ws, f'E{r}',
                f'=IFERROR(SUMPRODUCT(--($C${DF0}:$C${DF1}=$B{r}),'
                f'--($H${DF0}:$H${DF1}="Hecha"),$I${DF0}:$I${DF1}),"")',
                fmt=EUR)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Calendario 90 Días»
# --------------------------------------------------------------------------
def hoja_calendario(wb):
    ws = wb.create_sheet('Calendario 90 Días')
    cabecera(ws, 'Calendario a 90 días — 13 semanas, dos hitos por semana')
    apunte(ws, 'C3', 'Medir · escandallar y clasificar · aplicar · medir el '
                     'efecto · revisar. Los hitos son una propuesta: '
                     'reescríbelos con lo que toque en tu casa.')
    encabezados(ws, 4, [
        ('A', 'Semana', 9), ('B', 'Bloque', 28), ('C', 'Hito', 76),
        ('D', 'Responsable', 18), ('E', 'Fecha objetivo', 14),
        ('F', 'Hecho', 10), ('G', 'Nota', 32),
    ])
    ws.freeze_panes = 'C5'

    v_bloque, v_hecho, v_sem, v_resp = [], [], [], []
    for i in range(CF1 - CF0 + 1):
        r = CF0 + i
        if i < len(CALENDARIO):
            semana, bloq, hito, resp = CALENDARIO[i]
            motor.val(ws, f'A{r}', semana, fmt=ENT)
            motor.val(ws, f'B{r}', bloq)
            motor.val(ws, f'C{r}', hito, wrap=True)
            motor.val(ws, f'D{r}', resp)
            motor.val(ws, f'F{r}', 'No')
        else:
            ws[f'A{r}'].number_format = ENT
        motor.verde(ws, f'A{r}:D{r}')
        motor.verde(ws, f'F{r}:G{r}')
        v_sem.append(f'A{r}')
        v_bloque.append(f'B{r}')
        v_resp.append(f'D{r}')
        v_hecho.append(f'F{r}')
        motor.f(ws, f'E{r}',
                f"=IFERROR(IF(OR($A{r}=\"\",'Decisiones'!$D$36=\"\"),\"\","
                f"'Decisiones'!$D$36+7*($A{r}-1)),\"\")", fmt=FECHA)
    motor.dv_lista(ws, v_bloque, BLOQUES, titulo='Bloque no válido')
    motor.dv_lista(ws, v_resp, RESPONSABLES, titulo='Responsable no válido')
    motor.dv_lista(ws, v_hecho, SI_NO, titulo='Marca Sí o No')
    motor.dv_numerica(ws, v_sem, minimo=1, maximo=13, titulo='Semana del plan',
                      mensaje='El plan tiene 13 semanas: escribe un número '
                              'entre 1 y 13.')
    motor.semaforo_texto(ws, f'F{CF0}:F{CF1}', (
        ('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    bloque(ws, 'A32', 'RESUMEN — lo calcula el libro')
    motor.val(ws, 'C33', 'Hitos registrados')
    motor.f(ws, 'D33', f'=COUNTIF($C${CF0}:$C${CF1},"<>")', fmt=ENT)
    motor.val(ws, 'C34', 'Hitos hechos')
    motor.f(ws, 'D34', f'=COUNTIF($F${CF0}:$F${CF1},"Sí")', fmt=ENT)
    motor.val(ws, 'C35', 'Hitos pendientes')
    motor.f(ws, 'D35',
            '=IFERROR(IF(OR($D$33="",$D$34=""),"",$D$33-$D$34),"")', fmt=ENT)
    total(ws, 'C36', 'AVANCE DEL CALENDARIO (%)')
    total(ws, 'D36', '=IFERROR(IF($D$33=0,"",$D$34/$D$33),"")', fmt=PCT,
          formula=True)

    bloque(ws, 'A38', 'AVANCE POR BLOQUE — lo calcula el libro')
    encabezados(ws, 39, [
        ('C', 'Bloque', None), ('D', 'Hitos', None), ('E', 'Hechos', None),
        ('F', 'Avance (%)', None),
    ], alto=30)
    for k, bl in enumerate(BLOQUES):
        r = 40 + k
        motor.val(ws, f'C{r}', bl, bold=True)
        motor.f(ws, f'D{r}', f'=COUNTIF($B${CF0}:$B${CF1},$C{r})', fmt=ENT)
        motor.f(ws, f'E{r}',
                f'=IFERROR(SUMPRODUCT(--($B${CF0}:$B${CF1}=$C{r}),'
                f'--($F${CF0}:$F${CF1}="Sí")),"")', fmt=ENT)
        motor.f(ws, f'F{r}', f'=IFERROR(IF($D{r}=0,"",$E{r}/$D{r}),"")',
                fmt=PCT)
    motor.val(ws, 'C46',
              'Un plan recién abierto marca 0' + N + '% de avance: es lo '
              'correcto. El avance se gana marcando «Hecho» = Sí.')
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «KPI de Seguimiento»
# --------------------------------------------------------------------------
def hoja_kpi(wb):
    ws = wb.create_sheet('KPI de Seguimiento')
    cabecera(ws, 'KPI de seguimiento — mes 0 contra mes 3')
    apunte(ws, 'C3', 'El mes 0 es la foto de salida. En los KPI en porcentaje '
                     'la variación son PUNTOS porcentuales.')
    encabezados(ws, 4, [
        ('A', 'KPI', 42), ('B', 'Mes 0', 12), ('C', 'Mes 1', 12),
        ('D', 'Mes 2', 12), ('E', 'Mes 3', 12),
        ('F', 'Variación mes 1 vs mes 0', 14),
        ('G', 'Variación mes 2 vs mes 0', 14),
        ('H', 'Variación mes 3 vs mes 0', 14),
        ('I', '¿Bajar es bueno?', 13),
        ('J', 'Lectura mes 3 vs mes 0', 18), ('K', 'Nota', 52),
    ])
    ws.freeze_panes = 'B5'

    v_sino, v_num = [], []
    nombres = list(DE.KPI_SEGUIMIENTO.keys())
    for i in range(KF1 - KF0 + 1):
        r = KF0 + i
        if i < len(nombres):
            nombre = nombres[i]
            fmt, bajar, nota = KPI_META[nombre]
            motor.val(ws, f'A{r}', nombre)
            for k, col in enumerate('BCDE'):
                motor.val(ws, f'{col}{r}', DE.KPI_SEGUIMIENTO[nombre][k],
                          fmt=fmt)
            motor.val(ws, f'I{r}', bajar)
            motor.val(ws, f'K{r}', nota)
        else:
            fmt = '#,##0.00'
            for col in 'BCDE':
                ws[f'{col}{r}'].number_format = fmt
        for col in 'FGH':
            ws[f'{col}{r}'].number_format = fmt
        motor.verde(ws, f'A{r}:E{r}')
        motor.verde(ws, f'I{r}')
        motor.verde(ws, f'K{r}')
        v_sino.append(f'I{r}')
        v_num += [f'{c}{r}' for c in 'BCDE']
        for col_origen, col_var in (('C', 'F'), ('D', 'G'), ('E', 'H')):
            motor.f(ws, f'{col_var}{r}',
                    f'=IFERROR(IF(OR($B{r}="",${col_origen}{r}=""),"",'
                    f'${col_origen}{r}-$B{r}),"")', fmt=fmt)
        ws[f'H{r}'].font = Font(bold=True, color=AZUL)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF(OR($B{r}="",$E{r}="",$I{r}=""),"",'
                f'IF($H{r}=0,"Sin cambio",'
                f'IF(OR(AND($I{r}="Sí",$H{r}<0),AND($I{r}="No",$H{r}>0)),'
                f'"Mejora","Empeora"))),"")', bold=True)
    motor.dv_lista(ws, v_sino, SI_NO, titulo='Marca Sí o No',
                   mensaje='«Sí» si bajar es una mejora (food cost, prime '
                           'cost); «No» si lo bueno es subir (ticket medio, '
                           'margen).')
    motor.dv_numerica(ws, v_num, minimo=-1000000, titulo='Valor del KPI',
                      mensaje='Escribe el valor del KPI. Los porcentajes van '
                              'en tanto por uno: 0,30 = 30' + N + '%.')
    motor.semaforo_texto(ws, f'J{KF0}:J{KF1}', (
        ('Mejora', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Empeora', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('Sin cambio', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    bloque(ws, 'A14', 'RESUMEN — lo calcula el libro')
    motor.val(ws, 'A15', 'KPI registrados')
    motor.f(ws, 'C15', f'=COUNTIF($A${KF0}:$A${KF1},"<>")', fmt=ENT)
    total(ws, 'A16', 'KPI QUE MEJORAN')
    total(ws, 'C16', f'=COUNTIF($J${KF0}:$J${KF1},"Mejora")', fmt=ENT,
          formula=True)
    motor.val(ws, 'A17', 'KPI que empeoran')
    motor.f(ws, 'C17', f'=COUNTIF($J${KF0}:$J${KF1},"Empeora")', fmt=ENT)
    motor.val(ws, 'A18', 'KPI sin cambio')
    motor.f(ws, 'C18', f'=COUNTIF($J${KF0}:$J${KF1},"Sin cambio")', fmt=ENT)
    motor.val(ws, 'A19', 'KPI que mejoran (%)')
    motor.f(ws, 'C19', '=IFERROR(IF($C$15=0,"",$C$16/$C$15),"")', fmt=PCT)
    motor.regla_expresion(ws, 'C19', '=AND(ISNUMBER($C$19),$C$19<1)',
                          bg=motor.CF_AMBAR_BG, fg=motor.CF_AMBAR_FG)
    motor.val(ws, 'A21',
              'Los cuatro primeros KPI se leen juntos: un food cost que baja '
              'mientras el margen por cubierto también baja significa que has '
              'abaratado el plato, no que hayas ganado dinero.')
    motor.val(ws, 'A22',
              'Las tres últimas filas están libres para los KPI de tu casa '
              '(rotación de mesas, ratio de bebida sobre comida, mermas en '
              'euros…). Marca en «¿Bajar es bueno?» el sentido de cada uno.')
    setup(ws)
    return ws


# --------------------------------------------------------------------------
def mapa():
    return {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Decisiones': {
                'celdas': {
                    'Decisiones registradas': 'D27',
                    'Decisiones cerradas': 'D28',
                    'Decisiones descartadas': 'D29',
                    'Decisiones cerradas (%)': 'D30',
                    'Impacto total estimado (€/mes)': 'D31',
                    'Impacto ya conseguido (€/mes)': 'D32',
                    'Impacto pendiente (€/mes)': 'D33',
                    'Impacto total estimado a 12 meses (€)': 'D34',
                    'Objetivo de cierre a 90 días (%)': 'D35',
                    'Fecha de inicio del plan': 'D36',
                    'Reformular: decisiones': 'C39',
                    'Reformular: impacto estimado': 'D39',
                    'Resubir: decisiones': 'C40',
                    'Resubir: impacto estimado': 'D40',
                    'Rediseñar: decisiones': 'C41',
                    'Rediseñar: impacto estimado': 'D41',
                    'Retirar: decisiones': 'C42',
                    'Retirar: impacto estimado': 'D42',
                    'Negociar: decisiones': 'C43',
                    'Negociar: impacto estimado': 'D43',
                    'Mantener: decisiones': 'C44',
                    'Matriz multi-método: decisiones': 'C51',
                    'Matriz multi-método: impacto estimado': 'D51',
                    'Simulador multicanal: decisiones': 'C52',
                    'Simulador multicanal: impacto estimado': 'D52',
                    'Carta de bebidas: impacto estimado': 'D53',
                    'Cuadro de mando prime cost: impacto estimado': 'D54',
                    'Test de rendimiento: impacto estimado': 'D49',
                },
                'tablas': [
                    {'titulo': 'Decisiones: 10 de ejemplo y 10 filas libres',
                     'cols': [['#', 'A', 'num'],
                              ['Plato o área', 'B', 'txt'],
                              ['Herramienta de origen', 'C', 'txt'],
                              ['Decisión', 'D', 'txt'],
                              ['Responsable', 'E', 'txt'],
                              ['Semana del plan', 'F', 'num'],
                              ['Fecha objetivo', 'G', 'txt'],
                              ['Estado', 'H', 'txt'],
                              ['Impacto estimado (€/mes)', 'I', 'eur'],
                              ['Impacto ya conseguido (€/mes)', 'J', 'eur'],
                              ['Notas', 'K', 'txt']],
                     'filas': [5, 24]},
                    {'titulo': 'Resumen por tipo de decisión',
                     'cols': [['Decisión', 'B', 'txt'],
                              ['Decisiones', 'C', 'num'],
                              ['Impacto estimado (€/mes)', 'D', 'eur'],
                              ['Impacto conseguido (€/mes)', 'E', 'eur']],
                     'filas': [39, 44]},
                    {'titulo': 'Resumen por herramienta de origen',
                     'cols': [['Herramienta de origen', 'B', 'txt'],
                              ['Decisiones', 'C', 'num'],
                              ['Impacto estimado (€/mes)', 'D', 'eur'],
                              ['Impacto conseguido (€/mes)', 'E', 'eur']],
                     'filas': [48, 54]},
                ],
            },
            'Calendario 90 Días': {
                'celdas': {
                    'Hitos registrados': 'D33',
                    'Hitos hechos': 'D34',
                    'Hitos pendientes': 'D35',
                    'Avance del calendario (%)': 'D36',
                    'Medir: hitos': 'D40', 'Medir: avance (%)': 'F40',
                    'Escandallar y clasificar: hitos': 'D41',
                    'Escandallar y clasificar: avance (%)': 'F41',
                    'Aplicar decisiones: hitos': 'D42',
                    'Aplicar decisiones: avance (%)': 'F42',
                    'Medir el efecto: hitos': 'D43',
                    'Medir el efecto: avance (%)': 'F43',
                    'Revisión: hitos': 'D44', 'Revisión: avance (%)': 'F44',
                },
                'tablas': [
                    {'titulo': '13 semanas con dos hitos cada una',
                     'cols': [['Semana', 'A', 'num'], ['Bloque', 'B', 'txt'],
                              ['Hito', 'C', 'txt'],
                              ['Responsable', 'D', 'txt'],
                              ['Fecha objetivo', 'E', 'txt'],
                              ['Hecho', 'F', 'txt'], ['Nota', 'G', 'txt']],
                     'filas': [5, 30]},
                    {'titulo': 'Avance por bloque',
                     'cols': [['Bloque', 'C', 'txt'], ['Hitos', 'D', 'num'],
                              ['Hechos', 'E', 'num'],
                              ['Avance (%)', 'F', 'pct1']],
                     'filas': [40, 44]},
                ],
            },
            'KPI de Seguimiento': {
                'celdas': {
                    'Food cost: mes 0': 'B5', 'Food cost: mes 3': 'E5',
                    'Food cost: variación mes 3 vs mes 0': 'H5',
                    'Food cost: lectura': 'J5',
                    'Prime cost: mes 0': 'B6', 'Prime cost: mes 3': 'E6',
                    'Prime cost: variación mes 3 vs mes 0': 'H6',
                    'Prime cost: lectura': 'J6',
                    'Ticket medio: mes 0': 'B7', 'Ticket medio: mes 3': 'E7',
                    'Ticket medio: variación mes 3 vs mes 0': 'H7',
                    'MC por cubierto: mes 0': 'B8',
                    'MC por cubierto: mes 3': 'E8',
                    'MC por cubierto: variación mes 3 vs mes 0': 'H8',
                    'Platos en carta: mes 0': 'B9',
                    'Platos en carta: mes 3': 'E9',
                    'Platos en carta: variación mes 3 vs mes 0': 'H9',
                    'KPI registrados': 'C15', 'KPI que mejoran': 'C16',
                    'KPI que empeoran': 'C17', 'KPI sin cambio': 'C18',
                    'KPI que mejoran (%)': 'C19',
                },
                'tablas': [
                    {'titulo': 'KPI del trimestre: 5 sembrados y 3 libres',
                     'cols': [['KPI', 'A', 'txt'], ['Mes 0', 'B', 'num'],
                              ['Mes 1', 'C', 'num'], ['Mes 2', 'D', 'num'],
                              ['Mes 3', 'E', 'num'],
                              ['Variación mes 1 vs mes 0', 'F', 'num'],
                              ['Variación mes 2 vs mes 0', 'G', 'num'],
                              ['Variación mes 3 vs mes 0', 'H', 'num'],
                              ['¿Bajar es bueno?', 'I', 'txt'],
                              ['Lectura mes 3 vs mes 0', 'J', 'txt'],
                              ['Nota', 'K', 'txt']],
                     'filas': [5, 12]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_decisiones(wb)
    hoja_calendario(wb)
    hoja_kpi(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-22s %d' % (hoja, n))


if __name__ == '__main__':
    main()
