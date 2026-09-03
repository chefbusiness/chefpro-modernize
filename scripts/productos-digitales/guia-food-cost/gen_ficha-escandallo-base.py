#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_ficha-escandallo-base.py — libro 1 de los 8 del producto «Guía Food Cost
+ Ingeniería de Menú» (SPEC §2.2, fila 1).

Una hoja de ficha de escandallo AUTOSUFICIENTE (decisión D7 de la SPEC: sin el
bloque `Resumen`/`INDIRECT` de la maestra de la guía de 85 €) más la hoja de
Instrucciones de la familia.

Cifras: TODAS de `datos_ejemplo.FICHA` — no se teclea aquí ni un número de
negocio. Convenciones y helpers: `guias-v2_0/motor.py`.

Sale SIEMPRE en `build/ficha-escandallo-base.xlsx`. Después:
    python3 ../inject_cache.py build/ficha-escandallo-base.xlsx
"""
import io
import json
import os
import sys

AQUI = os.path.dirname(os.path.abspath(__file__))
GUIAS = '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0'
sys.path.insert(0, GUIAS)
sys.path.insert(0, AQUI)

from openpyxl import Workbook                                    # noqa: E402
from openpyxl.formatting.rule import FormulaRule                 # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill         # noqa: E402
from openpyxl.worksheet.page import PageMargins                  # noqa: E402
from openpyxl.worksheet.properties import PageSetupProperties    # noqa: E402

import motor                                                     # noqa: E402
import datos_ejemplo as D                                        # noqa: E402

# --------------------------------------------------------------------------
# Identidad del producto (NO se usa motor.version_line(): ésa dice «2.0 ·
# agosto 2026» y este producto es 1.0 de septiembre de 2026).
# --------------------------------------------------------------------------
PID = 'guia-food-cost-ingenieria-menu'
PRODUCTO = 'Guía Food Cost + Ingeniería de Menú'
NOMBRE = 'ficha-escandallo-base'
TITULO = 'Ficha de Escandallo Base'
SUBTITULO = 'AI Chef Pro · aichef.pro — Guía Food Cost + Ingeniería de Menú'
VERSION = ('Versión 1.0 · septiembre 2026 · aichef.pro/'
           + PID + ' · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 '
       '· johnguerrero.es')
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde, '
               'desprotege la hoja (sin contraseña).')
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

motor.CTX['producto'] = PID

CAB = '2D2D2D'          # cabecera de tabla, letra blanca
CREMA = 'FFF8DC'        # filas de total / bloques de resumen
FMT_EUR = motor.FMT_EUR
FMT_PCT = motor.FMT_PCT
FMT_CANT = '0.000'
FMT_ENT = '#,##0'

# Filas y columnas de la hoja «Ficha»
F_CAB = 9                 # cabecera de la tabla de ingredientes
F_INI = 10                # primera línea
N_LINEAS = 20
F_FIN = F_INI + N_LINEAS - 1          # 29
F_RES = F_FIN + 2                     # 31 · título del bloque RESUMEN

COLS = [
    # letra, cabecera, ancho, formato, entrada(verde)
    ('A', '#', 5, FMT_ENT, False),
    ('B', 'Ingrediente', 38, None, True),
    ('C', 'Unidad', 10, None, True),
    ('D', 'Cantidad NETA por ración', 15, FMT_CANT, True),
    ('E', 'Precio/Ud sin IVA (€)', 15, FMT_EUR, True),
    ('F', 'Merma (%)', 11, FMT_PCT, True),
    ('G', 'IVA de compra (%)', 13, FMT_PCT, True),
    ('H', 'Cantidad BRUTA a comprar', 15, FMT_CANT, False),
    ('I', 'Coste sin IVA (€)', 14, FMT_EUR, False),
    ('J', 'IVA soportado (€)', 14, FMT_EUR, False),
    ('K', 'Coste con IVA (€)', 14, FMT_EUR, False),
    ('L', 'Notas', 34, None, True),
]


# --------------------------------------------------------------------------
# Utilidades locales (el motor no las trae: son de presentación)
# --------------------------------------------------------------------------
def a4(ws, apaisado=True, titulos=None):
    """A4 completo: paperSize 9 + ajuste al ancho + pie (lo que mide el censo)."""
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.51, right=0.51, top=0.55, bottom=0.55,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def cabecera_libro(ws):
    motor.val(ws, 'A1', TITULO, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color='595959')


def fila_cabecera(ws, fila, columnas):
    for letra, texto, _, _, _ in columnas:
        cel = motor.val(ws, letra + str(fila), texto, bold=True, wrap=True)
        cel.fill = PatternFill('solid', fgColor=CAB)
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    ws.row_dimensions[fila].height = 32


def banda(ws, fila, desde, hasta, texto):
    """Banda CREMA de título de bloque."""
    motor.val(ws, desde + str(fila), texto, bold=True)
    for col in _letras(desde, hasta):
        ws[col + str(fila)].fill = PatternFill('solid', fgColor=CREMA)
        ws[col + str(fila)].font = Font(bold=True, size=11)


def _letras(desde, hasta):
    from openpyxl.utils import column_index_from_string, get_column_letter
    return [get_column_letter(i)
            for i in range(column_index_from_string(desde),
                           column_index_from_string(hasta) + 1)]


def semaforo_doble(ws, rango, ancla, umbral):
    """Rojo si supera el umbral, verde si lo cumple. Las dos con ISNUMBER."""
    motor.regla_expresion(
        ws, rango,
        '=AND(ISNUMBER(' + ancla + '),' + ancla + '>' + umbral + ')',
        bg=motor.CF_ROJO_BG, fg=motor.CF_ROJO_FG)
    ws.conditional_formatting.add(
        rango,
        FormulaRule(
            formula=['=AND(ISNUMBER(' + ancla + '),' + ancla + '<=' + umbral
                     + ')'],
            stopIfTrue=True,
            font=Font(color=motor.CF_VERDE_FG, bold=True),
            fill=PatternFill(start_color=motor.CF_VERDE_BG,
                             end_color=motor.CF_VERDE_BG, fill_type='solid')))


# --------------------------------------------------------------------------
# Hoja «Ficha»
# --------------------------------------------------------------------------
def hoja_ficha(wb):
    ws = wb.create_sheet('Ficha')
    cabecera_libro(ws)
    motor.anchos(ws, dict((c[0], c[2]) for c in COLS))

    # ---- datos del plato (celdas verdes de cabecera) ---------------------
    motor.val(ws, 'B4', 'Nombre del plato', bold=True)
    ws.merge_cells('D4:G4')
    motor.val(ws, 'D4', D.FICHA['plato'])
    motor.verde(ws, 'D4')

    motor.val(ws, 'B5', 'Familia de carta', bold=True)
    motor.val(ws, 'D5', D.FICHA['familia'])
    motor.verde(ws, 'D5')

    motor.val(ws, 'B6', 'Raciones que salen de esta ficha', bold=True)
    motor.val(ws, 'D6', D.FICHA['raciones'], fmt=FMT_ENT)
    motor.verde(ws, 'D6')

    motor.val(ws, 'B7', 'Food cost objetivo (%)', bold=True)
    motor.val(ws, 'D7', D.FICHA['food_cost_objetivo'], fmt=FMT_PCT)
    motor.verde(ws, 'D7')
    motor.val(ws, 'H7', 'El objetivo de la casa del ejemplo. Cámbialo y se '
                        'recalculan el PVP objetivo y el semáforo.', wrap=True)

    # ---- tabla de ingredientes ------------------------------------------
    fila_cabecera(ws, F_CAB, COLS)
    ws.freeze_panes = 'B' + str(F_INI)
    ws.auto_filter.ref = 'A' + str(F_CAB) + ':L' + str(F_FIN)

    lineas = D.FICHA['lineas']
    for i in range(N_LINEAS):
        r = F_INI + i
        motor.val(ws, 'A' + str(r), i + 1, fmt=FMT_ENT, align='center')
        if i < len(lineas):
            ing, uni, neta, precio, merma, iva = lineas[i]
            motor.val(ws, 'B' + str(r), ing)
            motor.val(ws, 'C' + str(r), uni, align='center')
            motor.val(ws, 'D' + str(r), neta, fmt=FMT_CANT)
            motor.val(ws, 'E' + str(r), precio, fmt=FMT_EUR)
            motor.val(ws, 'F' + str(r), merma, fmt=FMT_PCT)
            motor.val(ws, 'G' + str(r), iva, fmt=FMT_PCT)
        else:
            for letra, _, _, fmt, _ in COLS:
                if letra in ('B', 'C', 'D', 'E', 'F', 'G', 'L') and fmt:
                    ws[letra + str(r)].number_format = fmt
        for letra, _, _, _, entrada in COLS:
            if entrada:
                motor.verde(ws, letra + str(r))
        motor.f(ws, 'H' + str(r),
                '=IFERROR(IF(OR($D{r}="",$F{r}>=1),"",$D{r}/(1-$F{r})),"")'
                .format(r=r), fmt=FMT_CANT)
        motor.f(ws, 'I' + str(r),
                '=IFERROR(IF(OR($H{r}="",$E{r}=""),"",$H{r}*$E{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'J' + str(r),
                '=IFERROR(IF(OR($I{r}="",$G{r}=""),"",$I{r}*$G{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'K' + str(r),
                '=IFERROR(IF($I{r}="","",$I{r}+IF($J{r}="",0,$J{r})),"")'
                .format(r=r), fmt=FMT_EUR)

    # ---- bloque RESUMEN --------------------------------------------------
    r = F_RES
    banda(ws, r, 'B', 'G', 'RESUMEN — lo calcula el libro')

    r += 1
    F_TOTAL = r
    motor.val(ws, 'B' + str(r), 'COSTE TOTAL DE LA FICHA, sin IVA (€)',
              bold=True)
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($I${a}:$I${b},">0")=0,"",SUM($I${a}:$I${b}))'
            .format(a=F_INI, b=F_FIN), fmt=FMT_EUR, bold=True)

    r += 1
    F_RACION = r
    motor.val(ws, 'B' + str(r), 'Coste por ración, sin IVA (€)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${t}="",$D$6="",$D$6=0),"",$E${t}/$D$6),"")'
            .format(t=F_TOTAL), fmt=FMT_EUR, bold=True)

    r += 1
    F_PVP_OBJ = r
    motor.val(ws, 'B' + str(r),
              'PVP objetivo sin IVA (€) = coste por ración ÷ food cost '
              'objetivo')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${c}="",$D$7="",$D$7=0),"",$E${c}/$D$7),"")'
            .format(c=F_RACION), fmt=FMT_EUR)

    r += 1
    F_IVA = r
    motor.escribir_parametro(ws, r, 'B', 'E', 'iva_restauracion', col_nota='H')

    r += 1
    F_PVP_OBJ_IVA = r
    motor.val(ws, 'B' + str(r),
              'PVP objetivo CON IVA (€) — el que se imprime en la carta')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF($E${p}="","",$E${p}*(1+$E${i})),"")'
            .format(p=F_PVP_OBJ, i=F_IVA), fmt=FMT_EUR)

    r += 1
    F_PVP_ACT = r
    motor.val(ws, 'B' + str(r), 'PVP ACTUAL en carta, sin IVA (€)', bold=True)
    motor.val(ws, 'E' + str(r), D.FICHA['pvp_actual_sin_iva'], fmt=FMT_EUR)
    motor.verde(ws, 'E' + str(r))
    motor.val(ws, 'H' + str(r),
              'El precio que hoy tienes en la carta, quitándole el IVA. Es lo '
              'que compara el food cost REAL de abajo.', wrap=True)

    r += 1
    F_PVP_ACT_IVA = r
    motor.val(ws, 'B' + str(r), 'PVP actual CON IVA (€)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF($E${p}="","",$E${p}*(1+$E${i})),"")'
            .format(p=F_PVP_ACT, i=F_IVA), fmt=FMT_EUR)

    r += 1
    F_FC_REAL = r
    motor.val(ws, 'B' + str(r), 'Food cost REAL con el PVP actual (%)',
              bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${c}="",$E${p}="",$E${p}=0),"",$E${c}/$E${p}),"")'
            .format(c=F_RACION, p=F_PVP_ACT), fmt=FMT_PCT, bold=True)
    semaforo_doble(ws, 'E' + str(r), '$E$' + str(r), '$D$7')
    motor.val(ws, 'H' + str(r),
              'Verde si está en el objetivo de D7; rojo si lo supera.',
              wrap=True)

    r += 1
    F_MARGEN = r
    motor.val(ws, 'B' + str(r),
              'Margen de contribución con el PVP actual (€)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${c}="",$E${p}=""),"",$E${p}-$E${c}),"")'
            .format(c=F_RACION, p=F_PVP_ACT), fmt=FMT_EUR, bold=True)

    r += 1
    F_DIF = r
    motor.val(ws, 'B' + str(r),
              'Diferencia entre el PVP objetivo y el actual (€)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${o}="",$E${p}=""),"",$E${o}-$E${p}),"")'
            .format(o=F_PVP_OBJ, p=F_PVP_ACT), fmt=FMT_EUR)

    r += 1
    F_SUBIDA = r
    motor.val(ws, 'B' + str(r),
              'Subida necesaria sobre el PVP actual para llegar al objetivo (%)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${o}="",$E${p}="",$E${p}=0),"",$E${o}/$E${p}-1),"")'
            .format(o=F_PVP_OBJ, p=F_PVP_ACT), fmt=FMT_PCT)

    # ---- bloque IVA soportado (capítulo 04) ------------------------------
    r += 2
    banda(ws, r, 'B', 'G', 'IVA SOPORTADO — solo tesorería, NO es coste')

    r += 1
    F_IVA_TOT = r
    motor.val(ws, 'B' + str(r), 'IVA soportado total de la ficha (€)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($J${a}:$J${b},">0")=0,"",SUM($J${a}:$J${b}))'
            .format(a=F_INI, b=F_FIN), fmt=FMT_EUR)

    r += 1
    F_COSTE_IVA = r
    motor.val(ws, 'B' + str(r),
              'Coste con IVA soportado (solo tesorería) (€)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($K${a}:$K${b},">0")=0,"",SUM($K${a}:$K${b}))'
            .format(a=F_INI, b=F_FIN), fmt=FMT_EUR, bold=True)
    motor.val(ws, 'H' + str(r),
              'Lo que sale de la cuenta el día que pagas el albarán. El IVA '
              'soportado se deduce en el modelo 303: es tesorería, no coste, '
              'y por eso NO entra en el food cost.', wrap=True)

    r += 1
    F_FC_MAL = r
    motor.val(ws, 'B' + str(r),
              'Food cost si contases el IVA como coste (%) — NO se hace así')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${k}="",$D$6="",$D$6=0,$E${p}="",$E${p}=0),"",'
            '($E${k}/$D$6)/$E${p}),"")'.format(k=F_COSTE_IVA, p=F_PVP_ACT),
            fmt=FMT_PCT)
    motor.val(ws, 'H' + str(r),
              'Está aquí para que veas el error: contar el IVA soportado como '
              'coste infla el food cost y te empuja a subir precios que no '
              'hacen falta.', wrap=True)

    # ---- validación de datos --------------------------------------------
    filas = [str(x) for x in range(F_INI, F_FIN + 1)]
    motor.dv_lista(ws, ['C' + x for x in filas],
                   ['kg', 'L', 'ud', 'g', 'cl', 'ración', 'docena'],
                   titulo='Unidad no válida')
    motor.dv_lista(ws, ['D5'],
                   ['Entrantes', 'Principales', 'Postres', 'Menú', 'Bebidas'],
                   titulo='Familia no válida')
    motor.dv_numerica(ws, ['D' + x for x in filas], minimo=0,
                      titulo='Cantidad neta',
                      prompt='La cantidad que se SIRVE en el plato, ya limpia '
                             'y en la unidad de compra (0,220 kg).')
    motor.dv_numerica(ws, ['E' + x for x in filas], minimo=0,
                      titulo='Precio de compra',
                      prompt='Precio por unidad de compra SIN IVA, el de tu '
                             'albarán.')
    motor.dv_porcentaje(ws, ['F' + x for x in filas],
                        titulo=motor.PCT_MERMA[0], prompt=motor.PCT_MERMA[1],
                        maximo=motor.PCT_MERMA[2])
    # Lista cerrada, no un rango 0-21 %: sólo existen tres tipos de IVA
    # soportado en España (hoja «Albarán e IVA soportado»). Un rango decimal
    # aceptaba 0,07 o 0,15 sin avisar, tipos que no existen en la Ley del IVA.
    motor.dv_lista(ws, ['G' + x for x in filas], ['0.04', '0.1', '0.21'],
                   titulo='IVA de compra no válido',
                   mensaje='El IVA soportado en España sólo tiene tres tipos: '
                           '4 %, 10 % o 21 % (hoja «Albarán e IVA soportado»). '
                           'Es informativo: se deduce en el 303 y no entra en '
                           'el coste.')
    motor.dv_numerica(ws, ['D6'], minimo=1, titulo='Raciones',
                      prompt='Cuántas raciones salen de las cantidades de esta '
                             'ficha.')
    motor.dv_porcentaje(ws, ['D7'], titulo='Food cost objetivo',
                        prompt='Se escribe en tanto por uno: 0,30 = 30 %.')
    motor.dv_porcentaje(ws, ['E' + str(F_IVA)],
                        titulo='Tipo de IVA', maximo=0.21,
                        prompt='Tipo de IVA repercutido del canal. 0,10 en '
                               'sala.')
    motor.dv_numerica(ws, ['E' + str(F_PVP_ACT)], minimo=0,
                      titulo='PVP actual',
                      prompt='El precio de tu carta SIN IVA.')

    a4(ws, apaisado=True, titulos=str(F_CAB) + ':' + str(F_CAB))

    return ws, {
        'total': 'E' + str(F_TOTAL), 'racion': 'E' + str(F_RACION),
        'pvp_obj': 'E' + str(F_PVP_OBJ), 'iva': 'E' + str(F_IVA),
        'pvp_obj_iva': 'E' + str(F_PVP_OBJ_IVA),
        'pvp_act': 'E' + str(F_PVP_ACT), 'pvp_act_iva': 'E' + str(F_PVP_ACT_IVA),
        'fc_real': 'E' + str(F_FC_REAL), 'margen': 'E' + str(F_MARGEN),
        'dif': 'E' + str(F_DIF), 'subida': 'E' + str(F_SUBIDA),
        'iva_tot': 'E' + str(F_IVA_TOT), 'coste_iva': 'E' + str(F_COSTE_IVA),
        'fc_mal': 'E' + str(F_FC_MAL),
    }


# --------------------------------------------------------------------------
# Hoja «Albarán e IVA soportado» (capítulo 04: el coste real de compra, 4 %,
# 10 % y 21 % en el mismo albarán). Antes de esta hoja los tres tipos sólo
# vivían en el PROMPT de la validación de G10:G29 y el capítulo no tenía
# ninguna celda que citar (SPEC §7-bis.7: el PDF cita celdas, no inventa).
# --------------------------------------------------------------------------
AL_TIPOS_CAB = 4
AL_TIPOS_INI = 5
AL_TIPOS_FIN = AL_TIPOS_INI + len(D.IVA_SOPORTADO) - 1     # 7
AL_CAB = AL_TIPOS_FIN + 3
AL_INI = AL_CAB + 1
AL_FIN = AL_INI + len(D.ALBARAN_EJEMPLO) - 1
AL_TOT = AL_FIN + 1


def hoja_albaran(wb):
    ws = wb.create_sheet('Albarán e IVA soportado')
    cabecera_libro(ws)
    motor.anchos(ws, {'A': 5, 'B': 32, 'C': 11, 'D': 10, 'E': 15, 'F': 15,
                      'G': 12, 'H': 13, 'I': 14})

    banda(ws, 3, 'A', 'D', 'LOS TRES TIPOS DE IVA SOPORTADO EN COMPRAS')
    fila_cabecera(ws, AL_TIPOS_CAB,
                  [('A', 'Tipo de IVA soportado', 15, FMT_PCT, True),
                   ('B', 'Qué compra lleva este tipo', 70, None, False)])
    for i, (tipo, nota) in enumerate(D.IVA_SOPORTADO):
        r = AL_TIPOS_INI + i
        motor.val(ws, 'A' + str(r), tipo, fmt=FMT_PCT)
        motor.verde(ws, 'A' + str(r))
        motor.val(ws, 'B' + str(r), nota, wrap=True)
        ws.row_dimensions[r].height = 44
    motor.val(ws, 'A' + str(AL_TIPOS_FIN + 1), 'Fuente: ' + D.FUENTE_IVA,
              wrap=True)
    motor.dv_lista(ws, ['A' + str(r) for r in
                        range(AL_TIPOS_INI, AL_TIPOS_FIN + 1)],
                   ['0.04', '0.1', '0.21'], titulo='IVA soportado no válido')

    banda(ws, AL_TIPOS_FIN + 3, 'A', 'D',
          'ALBARÁN DE EJEMPLO — CAPÍTULO 04')
    fila_cabecera(ws, AL_CAB, [
        ('A', '#', 5, FMT_ENT, False), ('B', 'Producto', 32, None, True),
        ('C', 'Cantidad', 11, FMT_CANT, True),
        ('D', 'Unidad', 10, None, True),
        ('E', 'Precio/Ud sin IVA (€)', 15, FMT_EUR, True),
        ('F', 'Base imponible (€)', 15, FMT_EUR, False),
        ('G', 'Tipo de IVA (%)', 12, FMT_PCT, True),
        ('H', 'Cuota de IVA (€)', 13, FMT_EUR, False),
        ('I', 'Total con IVA (€)', 14, FMT_EUR, False)])
    ws.freeze_panes = 'B' + str(AL_INI)
    for i, (nombre, cantidad, unidad, precio, tipo) in enumerate(
            D.ALBARAN_EJEMPLO):
        r = AL_INI + i
        motor.val(ws, 'A' + str(r), i + 1, fmt=FMT_ENT, align='center')
        motor.val(ws, 'B' + str(r), nombre, verde_=True)
        motor.val(ws, 'C' + str(r), cantidad, fmt=FMT_CANT, verde_=True)
        motor.val(ws, 'D' + str(r), unidad, verde_=True)
        motor.val(ws, 'E' + str(r), precio, fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'G' + str(r), tipo, fmt=FMT_PCT, verde_=True)
        motor.f(ws, 'F' + str(r),
                '=IFERROR(IF(OR($C{r}="",$E{r}=""),"",$C{r}*$E{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'H' + str(r),
                '=IFERROR(IF(OR($F{r}="",$G{r}=""),"",$F{r}*$G{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'I' + str(r),
                '=IFERROR(IF($F{r}="","",$F{r}+IF($H{r}="",0,$H{r})),"")'
                .format(r=r), fmt=FMT_EUR)
    motor.val(ws, 'B' + str(AL_TOT), 'TOTAL ALBARÁN', bold=True)
    for col in ('F', 'H', 'I'):
        motor.f(ws, col + str(AL_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=AL_INI, b=AL_FIN), fmt=FMT_EUR, bold=True)
    for col in 'ABCDEFGHI':
        ws[col + str(AL_TOT)].fill = PatternFill('solid', fgColor=CREMA)
        ws[col + str(AL_TOT)].font = Font(bold=True, size=10)
    motor.val(ws, 'B' + str(AL_TOT + 2),
              'El mismo albarán trae los tres tipos a la vez: aceite de '
              'oliva, tomate y queso curado al 4 %; solomillo al 10 %; vino '
              'y envases al 21 %. Cuadra la columna «Tipo de IVA (%)» de tu '
              'ficha con esta tabla antes de escandallar.', wrap=True)

    filas_al = [str(x) for x in range(AL_INI, AL_FIN + 1)]
    motor.dv_lista(ws, ['G' + x for x in filas_al], ['0.04', '0.1', '0.21'],
                   titulo='IVA de compra no válido')
    motor.dv_numerica(ws, ['C' + x for x in filas_al], minimo=0,
                      titulo='Cantidad')
    motor.dv_numerica(ws, ['E' + x for x in filas_al], minimo=0,
                      titulo='Precio de compra')

    a4(ws, apaisado=True)
    return {
        'tipo_4': 'A' + str(AL_TIPOS_INI), 'tipo_10': 'A' + str(AL_TIPOS_INI + 1),
        'tipo_21': 'A' + str(AL_TIPOS_INI + 2),
        'albaran_total_base': 'F' + str(AL_TOT),
        'albaran_total_iva': 'H' + str(AL_TOT),
        'albaran_total_con_iva': 'I' + str(AL_TOT),
    }


# --------------------------------------------------------------------------
# Hoja «Instrucciones» (primera)
# --------------------------------------------------------------------------
PASOS = [
    '1. Escribe arriba el nombre del plato, la familia, las raciones que salen '
    'de la ficha y tu food cost objetivo.',
    '2. Rellena una línea por ingrediente: cantidad NETA por ración (la que se '
    'sirve), precio de compra sin IVA y merma en tanto por uno (0,12 = 12 %).',
    '3. El libro calcula la cantidad BRUTA que hay que comprar con la fórmula '
    'cantidad neta / (1 - merma), y de ahí el coste de cada línea.',
    '4. La columna «IVA de compra (%)» es informativa: el IVA soportado se '
    'deduce en el modelo 303, así que no entra en el coste ni en el food '
    'cost. La hoja «Albarán e IVA soportado» trae los tres tipos (4 %, '
    '10 %, 21 %) con qué producto lleva cada uno y un albarán de ejemplo '
    'con los tres tipos mezclados.',
    '5. En el bloque RESUMEN tienes el coste por ración, el PVP objetivo sin '
    'IVA y con IVA, y el food cost real del precio que hoy tienes en la carta.',
    '6. Si el food cost real sale en rojo, tu precio actual está por encima del '
    'objetivo: sube el precio, baja el coste o revisa el gramaje.',
    '7. Duplica la hoja «Ficha» (clic derecho en la pestaña, Mover o copiar, '
    'Crear una copia) para tener una ficha por plato.',
]

NOTA_IVA_LIBRO = (
    'El tipo de IVA de este libro es el de SALA: 10 % para la comida y la '
    'bebida consumidas en el local, alcohol incluido (art. 91.Uno.2.2.º de la '
    'Ley 37/1992). Para llevar y en delivery el tipo depende del producto: '
    'cámbialo en la celda verde «Tipo de IVA de restauración (%)».')


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 108
    motor.val(ws, 'A1', TITULO, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color='595959')
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    for texto in (motor.NOTA_VERDES, NOTA_IVA_LIBRO, DESPROTEGER, BIO,
                  VERSION):
        cel = motor.val(ws, 'A' + str(fila), texto, wrap=True)
        if texto is motor.NOTA_VERDES:
            cel.fill = PatternFill('solid', fgColor=motor.VERDE)
        fila += 1
    a4(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Mapa de celdas: el contrato con el guion de la guía (SPEC §4). Las «celdas»
# son RESULTADOS citables; las «tablas», rangos imprimibles con cabecera.
# --------------------------------------------------------------------------
TIPO_POR_FMT = {FMT_EUR: 'eur', FMT_PCT: 'pct1', FMT_CANT: 'num1',
                FMT_ENT: 'num', None: 'txt'}


def escribir_mapa(celdas, celdas_albaran):
    cols = [[c[1], c[0], TIPO_POR_FMT[c[3]]] for c in COLS]
    mapa = {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Instrucciones': {
                'celdas': {
                    'Título del libro': 'A1',
                    'Nota de IVA de sala': 'A14',
                    'Línea de versión': 'A17',
                },
                'tablas': [],
            },
            'Ficha': {
                'celdas': {
                    'Nombre del plato': 'D4',
                    'Familia de carta': 'D5',
                    'Raciones de la ficha': 'D6',
                    'Food cost objetivo': 'D7',
                    'Coste total de la ficha sin IVA': celdas['total'],
                    'Coste por ración sin IVA': celdas['racion'],
                    'PVP objetivo sin IVA': celdas['pvp_obj'],
                    'Tipo de IVA de restauración (sala)': celdas['iva'],
                    'PVP objetivo con IVA': celdas['pvp_obj_iva'],
                    'PVP actual en carta sin IVA': celdas['pvp_act'],
                    'PVP actual con IVA': celdas['pvp_act_iva'],
                    'Food cost real con el PVP actual': celdas['fc_real'],
                    'Margen de contribución con el PVP actual': celdas['margen'],
                    'Diferencia entre PVP objetivo y actual': celdas['dif'],
                    'Subida necesaria para llegar al objetivo': celdas['subida'],
                    'IVA soportado total de la ficha': celdas['iva_tot'],
                    'Coste con IVA soportado (solo tesorería)': celdas['coste_iva'],
                    'Food cost erróneo contando el IVA como coste': celdas['fc_mal'],
                    # Anclas de línea que citan los capítulos 04, 05 y 06.
                    'Cantidad bruta de la línea 1 (solomillo)':
                        'H' + str(F_INI),
                    'Coste sin IVA de la línea 1 (solomillo)':
                        'I' + str(F_INI),
                    'Coste con IVA de la línea 1 (solomillo)':
                        'K' + str(F_INI),
                    'IVA soportado de la línea 7 (vino Pedro Ximénez)':
                        'J' + str(F_INI + 6),
                    'Cantidad bruta de la línea 3 (boniato, merma 18 %)':
                        'H' + str(F_INI + 2),
                },
                # Solo van aquí los rangos HOMOGÉNEOS con cabecera. Los
                # bloques etiqueta/valor del RESUMEN mezclan € y %, así que
                # se citan celda a celda desde «celdas».
                'tablas': [
                    {'titulo': 'Líneas de la ficha de escandallo',
                     'cols': cols,
                     'filas': [F_INI, F_FIN]},
                ],
            },
            'Albarán e IVA soportado': {
                'celdas': {
                    'Tipo de IVA soportado al 4 %': celdas_albaran['tipo_4'],
                    'Tipo de IVA soportado al 10 %':
                        celdas_albaran['tipo_10'],
                    'Tipo de IVA soportado al 21 %':
                        celdas_albaran['tipo_21'],
                    'Base imponible total del albarán de ejemplo':
                        celdas_albaran['albaran_total_base'],
                    'Cuota de IVA total del albarán de ejemplo':
                        celdas_albaran['albaran_total_iva'],
                    'Total con IVA del albarán de ejemplo':
                        celdas_albaran['albaran_total_con_iva'],
                },
                'tablas': [
                    {'titulo': 'Los tres tipos de IVA soportado en compras',
                     'cols': [['Tipo de IVA soportado', 'A', 'pct1'],
                              ['Qué compra lleva este tipo', 'B', 'txt']],
                     'filas': [AL_TIPOS_INI, AL_TIPOS_FIN]},
                    {'titulo': 'Albarán de ejemplo con los tres tipos de IVA',
                     'cols': [['#', 'A', 'num'], ['Producto', 'B', 'txt'],
                              ['Cantidad', 'C', 'num1'],
                              ['Unidad', 'D', 'txt'],
                              ['Precio/Ud sin IVA (€)', 'E', 'eur'],
                              ['Base imponible (€)', 'F', 'eur'],
                              ['Tipo de IVA (%)', 'G', 'pct1'],
                              ['Cuota de IVA (€)', 'H', 'eur'],
                              ['Total con IVA (€)', 'I', 'eur']],
                     'filas': [AL_INI, AL_FIN]},
                ],
            },
        },
    }
    destino = os.path.join(AQUI, 'build', 'mapa-' + NOMBRE + '.json')
    with io.open(destino, 'w', encoding='utf-8') as fh:
        fh.write(json.dumps(mapa, ensure_ascii=False, indent=2,
                            sort_keys=False))
        fh.write(u'\n')
    return destino


# --------------------------------------------------------------------------
def main():
    motor.REGISTRO = []
    wb = Workbook()
    wb.remove(wb.active)
    ws_ficha, celdas = hoja_ficha(wb)
    celdas_albaran = hoja_albaran(wb)
    hoja_instrucciones(wb)
    wb.move_sheet('Instrucciones', offset=-2)
    wb.active = 0

    motor.normalizar_texto(wb, NOMBRE + '.xlsx')
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · Versión 1.0 · septiembre 2026'

    destino = os.path.join(AQUI, 'build', NOMBRE + '.xlsx')
    if not os.path.isdir(os.path.dirname(destino)):
        os.makedirs(os.path.dirname(destino))
    wb.save(destino)
    mapa = escribir_mapa(celdas, celdas_albaran)
    print('escrito: ' + destino)
    print('mapa:    ' + mapa)
    print('formulas registradas: ' + str(len(motor.REGISTRO)))
    print('celdas de resultado citables: ' + str(len(celdas)))
    return destino


if __name__ == '__main__':
    main()
