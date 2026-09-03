#!/usr/bin/env python3
"""
gen_carta-de-bebidas-beverage-cost.py — libro 6 de «Guía Food Cost + Ingeniería
de Menú» (SPEC §2.2 fila 6, decisiones D4 y D12).

Hojas: Instrucciones · Parámetros · Vinos · Cervezas y Refrescos ·
Destilados y Cócteles · Resumen Bodega.

La bodega como cuenta de resultados propia: coste por copa y por servicio,
margen, beverage cost por referencia y ponderado por categoría, y el PVP con
IVA correcto en sala y para llevar (la matriz 3×3 de la decisión D4).

Convenciones de la familia (motor.py de guias-v2_0): celdas verdes = entrada;
ninguna celda con fórmula en verde; cero constantes dentro de una fórmula (IVA
y objetivos viven en celdas); «sin dato» = "" nunca 0; IFERROR en todo
cociente; semáforos con ISNUMBER; prohibidas INDIRECT, COUNTA, PMT, OFFSET,
XLOOKUP, LET, LAMBDA y matrices dinámicas.

Dos decisiones de modelado que el juego de datos obligaba a tomar y que quedan
DICHAS en el propio libro:
  · El vermut de grifo se compra en botella de 1 L, no de 70 cl. Lo confirma el
    propio juego de datos: el negroni valora el «Vermut rojo» a 7,90 €/L y la
    botella cuesta 7,90 €. Por eso el formato de la botella es una celda verde
    por fila y no una constante de 70 cl dentro de la fórmula.
  · El gin tonic aparece en la tabla de destilados (costeado como destilado +
    mezcla) y en la de cócteles (desglosado ingrediente a ingrediente). Son el
    MISMO servicio contado de dos maneras, así que la columna «¿Suma al total
    de la bodega?» lo deja fuera del total para no contarlo dos veces.

Salida: build/carta-de-bebidas-beverage-cost.xlsx
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
sys.path.insert(0, AQUI)

import motor                                                   # noqa: E402
import datos_ejemplo as D                                      # noqa: E402

motor.CTX['producto'] = 'guia-food-cost-ingenieria-menu'

PRODUCTO = 'Guía Food Cost + Ingeniería de Menú'
SUBTITULO = 'AI Chef Pro · aichef.pro — ' + PRODUCTO
VERSION = ('Versión 1.0 · septiembre 2026 · '
           'aichef.pro/guia-food-cost-ingenieria-menu · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 '
       '· johnguerrero.es')
NOTA_DESPROTEGER = ('Para editar la estructura o una celda que no esté en '
                    'verde, desprotege la hoja (sin contraseña).')
TITULO = 'Carta de bebidas y beverage cost'
NOMBRE = 'carta-de-bebidas-beverage-cost'

FMT_EUR = '#,##0.00 €'
FMT_PCT = '0.0%'
FMT_ENT = '#,##0'
FMT_CANT = '0.000'

GRIS = 'F2F2F2'
CREMA = 'FFF6DC'
ORO = 'FFD700'
CABECERA = '2D2D2D'

CANALES = ('Sala', 'Take away', 'Delivery')
TIPOS = ('Comida', 'Refresco/azucarada', 'Bebida alcohólica')

# Ancla de canal para el INDEX/MATCH de la matriz de IVA (fila 6 = Sala,
# fila 7 = Take away). Se referencia la CELDA, nunca el texto literal.
ANCLA_SALA = 'Parámetros!$A$6'
ANCLA_LLEVAR = 'Parámetros!$A$7'
MATRIZ = 'Parámetros!$B$6:$D$8'
CANALES_RG = 'Parámetros!$A$6:$A$8'
TIPOS_RG = 'Parámetros!$B$5:$D$5'


# --------------------------------------------------------------------------
# Utilidades de estilo
# --------------------------------------------------------------------------
def cabecera(ws, fila, columnas, altura=44):
    for letra, texto in columnas:
        c = ws[letra + str(fila)]
        c.value = texto
        c.fill = PatternFill('solid', fgColor=CABECERA)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    ws.row_dimensions[fila].height = altura


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12)


def fila_total(ws, fila, primera, ultima):
    for col in range(motor.column_index_from_string(primera),
                     motor.column_index_from_string(ultima) + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = PatternFill('solid', fgColor=CREMA)
        c.font = Font(bold=True)


def cf_expresion(ws, rango, formula, bg, fg):
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


def semaforo_objetivo(ws, rango, ancla_col, ancla_obj, fila_ini):
    """Rojo si el beverage cost supera el objetivo, verde si no (ISNUMBER)."""
    cf_expresion(ws, rango,
                 '=AND(ISNUMBER(${c}{f}),ISNUMBER({o}),${c}{f}>{o})'
                 .format(c=ancla_col, f=fila_ini, o=ancla_obj),
                 motor.CF_ROJO_BG, motor.CF_ROJO_FG)
    cf_expresion(ws, rango,
                 '=AND(ISNUMBER(${c}{f}),ISNUMBER({o}),${c}{f}<={o})'
                 .format(c=ancla_col, f=fila_ini, o=ancla_obj),
                 motor.CF_VERDE_BG, motor.CF_VERDE_FG)


def pagina(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezar(ws, titulo, nota=None):
    motor.val(ws, 'A1', titulo)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    if nota:
        motor.val(ws, 'A3', nota)
        ws['A3'].font = Font(italic=True, size=9)


def anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


def objetivo_mirror(ws, fila, celda_origen, etiqueta):
    """Espeja el objetivo de la categoría EN LA MISMA HOJA.

    El formato condicional no puede referirse a otra hoja en Google Sheets y la
    única forma de hacerlo sería INDIRECT, que está prohibida.
    """
    motor.val(ws, 'B' + str(fila), etiqueta, bold=True)
    motor.f(ws, 'E' + str(fila),
            '=IF({0}="","",{0})'.format(celda_origen), fmt=FMT_PCT)
    ws['E' + str(fila)].fill = PatternFill('solid', fgColor=GRIS)
    return '$E$' + str(fila)


def iva_sala(fila, col_tipo='C'):
    return ('=IFERROR(IF(${c}{f}="","",INDEX({m},MATCH({a},{cr},0),'
            'MATCH(${c}{f},{tr},0))),"")'
            .format(c=col_tipo, f=fila, m=MATRIZ, a=ANCLA_SALA, cr=CANALES_RG,
                    tr=TIPOS_RG))


def iva_llevar(fila, col_tipo='C'):
    return ('=IFERROR(IF(${c}{f}="","",INDEX({m},MATCH({a},{cr},0),'
            'MATCH(${c}{f},{tr},0))),"")'
            .format(c=col_tipo, f=fila, m=MATRIZ, a=ANCLA_LLEVAR,
                    cr=CANALES_RG, tr=TIPOS_RG))


# --------------------------------------------------------------------------
# Instrucciones
# --------------------------------------------------------------------------
PASOS = [
    '1. Hoja «Parámetros»: revisa la matriz de IVA repercutido y pon tu '
    'objetivo de beverage cost para cada categoría. Los tres objetivos vienen '
    'sembrados con las referencias del sector y su fuente.',
    '2. Hoja «Vinos»: por cada referencia, el precio de compra de la botella '
    'sin IVA, el formato, los centilitros que sirves por copa y los PVP de '
    'botella y de copa. El libro calcula las copas por botella, el coste por '
    'copa y el beverage cost de las dos formas de venderlo. Hay 30 filas '
    '(8 de ejemplo, 22 libres): una carta de vinos real tiene 30-60 '
    'referencias.',
    '3. Hoja «Cervezas y Refrescos»: el coste por servicio sale del precio de '
    'la unidad de compra por la parte que sirves (barril de 30 L → caña de 25 '
    'cl). La «unidad de medida» dice si el contenido y el servicio van en cl '
    'o en ml. Hay 15 filas (7 de ejemplo, 8 libres).',
    '4. Hoja «Destilados y Cócteles»: arriba, el combinado sencillo (destilado '
    '+ mezcla, 12 filas: 5 de ejemplo y 7 libres); abajo, los cócteles '
    'desglosados ingrediente a ingrediente (8 cócteles con hasta 4 líneas '
    'cada uno: 4 de ejemplo y 4 libres), con el precio de cada uno en euros '
    'por litro.',
    '5. Escribe las unidades vendidas al mes de cada referencia: sin ellas hay '
    'coste por copa, pero no hay beverage cost ponderado ni margen del mes.',
    '6. Hoja «Resumen Bodega»: ventas, coste, beverage cost ponderado y margen '
    'de contribución por categoría, con semáforo contra tu objetivo y el total '
    'de la bodega.',
    '7. Los precios con IVA se calculan solos por canal: en sala todo va al '
    '10 %; para llevar, el alcohol y los refrescos azucarados al 21 %.',
]

NOTAS_LIBRO = [
    'Todos los precios de compra y de venta van SIN IVA salvo las columnas que '
    'dicen «con IVA»: son las que van en la carta.',
    'El IVA repercutido depende del canal Y del tipo de producto. En sala todo '
    'el consumo va al 10 %, alcohol incluido (art. 91.Uno.2.2.º de la Ley del '
    'IVA). Para llevar es entrega de bienes: las bebidas alcohólicas y los '
    'refrescos, zumos y gaseosas con azúcares o edulcorantes añadidos van al '
    '21 %; el agua, los zumos naturales y la cerveza sin alcohol se quedan en '
    'el tipo reducido, que es la casilla «Comida» de la matriz.',
    'El beverage cost se mide sobre la venta NETA (base imponible) y con el '
    'coste NETO de IVA soportado: el IVA de las compras se deduce en el modelo '
    '303, es tesorería, no coste.',
    'La bodega no se gestiona con la media: se gestiona referencia a '
    'referencia. Un vino de la casa al 29 % y un albariño al 35 % pueden dar la '
    'misma media que dos vinos al 32 % y no significan lo mismo.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    anchos(ws, {'A': 110.0})
    motor.val(ws, 'A1', TITULO)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    seccion(ws, 'A4', 'Instrucciones de uso')
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), 'Celdas verdes = campos editables')
    ws['A' + str(fila)].fill = PatternFill('solid', fgColor=motor.VERDE)
    fila += 2
    for nota in NOTAS_LIBRO:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), NOTA_DESPROTEGER)
    fila += 1
    motor.val(ws, 'A' + str(fila), BIO)
    fila += 1
    motor.val(ws, 'A' + str(fila), VERSION)
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Parámetros
# --------------------------------------------------------------------------
def hoja_parametros(wb):
    ws = wb.create_sheet('Parámetros')
    anchos(ws, {'A': 26, 'B': 22, 'C': 22, 'D': 24, 'E': 20})
    encabezar(ws, 'Parámetros de la bodega')

    seccion(ws, 'A4', 'MATRIZ DE IVA REPERCUTIDO POR CANAL Y TIPO DE PRODUCTO')
    cabecera(ws, 5, [('A', 'Canal'), ('B', TIPOS[0]), ('C', TIPOS[1]),
                     ('D', TIPOS[2])], altura=32)
    for i, canal in enumerate(CANALES):
        fila = 6 + i
        motor.val(ws, 'A' + str(fila), canal, bold=True)
        for j, tipo in enumerate(TIPOS):
            motor.val(ws, chr(ord('B') + j) + str(fila),
                      D.IVA_REPERCUTIDO[canal][tipo], fmt=FMT_PCT, verde_=True)
    fila = 9
    for canal in CANALES:
        motor.val(ws, 'A' + str(fila), canal + ' — ' + D.NOTAS_IVA[canal])
        fila += 1
    motor.val(ws, 'A12', 'Fuente: ' + D.FUENTE_IVA)
    motor.val(ws, 'A13',
              'El agua, los zumos naturales sin azúcares añadidos y la cerveza '
              'sin alcohol no son bebidas alcohólicas ni llevan edulcorantes '
              'añadidos: para llevar van al tipo reducido, es decir, a la '
              'columna «Comida» de esta matriz.')

    seccion(ws, 'A15', 'OBJETIVO DE BEVERAGE COST POR CATEGORÍA')
    cabecera(ws, 16, [('A', 'Categoría'), ('B', 'Objetivo (%)')], altura=20)
    for i, (categoria, objetivo, nota) in enumerate(D.BEVERAGE_COST_REFERENCIA):
        fila = 17 + i
        motor.val(ws, 'A' + str(fila), categoria, bold=True)
        motor.val(ws, 'B' + str(fila), objetivo, fmt=FMT_PCT, verde_=True)
    fila = 20
    for categoria, objetivo, nota in D.BEVERAGE_COST_REFERENCIA:
        motor.val(ws, 'A' + str(fila), categoria + ' — ' + nota)
        fila += 1
    motor.val(ws, 'A23', 'Fuente: ' + D.FUENTE_BEVERAGE)
    motor.val(ws, 'A24',
              'Son objetivos de la casa, no una norma: súbelos o bájalos según '
              'tu carta y tu clientela. Todo el libro los lee de estas tres '
              'celdas.')

    motor.dv_porcentaje(ws, ['%s%d' % (c, r) for c in 'BCD'
                             for r in (6, 7, 8)],
                        titulo='Tipo de IVA',
                        prompt='Se escribe en tanto por uno: 0,10 = 10 %.')
    motor.dv_porcentaje(ws, ['B17', 'B18', 'B19'],
                        titulo='Beverage cost objetivo',
                        prompt='Se escribe en tanto por uno: 0,30 = 30 %.',
                        maximo=0.95)
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Vinos
# --------------------------------------------------------------------------
V_CAB, V_INI = 4, 5
CAP_VINOS = 30            # capacidad de la hoja: una carta de vinos real tiene
                          # 30-60 referencias, no los 8 sembrados de ejemplo
V_FIN = V_INI + CAP_VINOS - 1
V_TOT = V_FIN + 1
V_OBJ = V_TOT + 2


def hoja_vinos(wb):
    ws = wb.create_sheet('Vinos')
    encabezar(ws, 'Vinos: botella y copa',
              nota='La misma botella deja un margen distinto según la vendas '
                   'entera o por copas. Aquí se ven las dos.')
    cols = [('A', '#'), ('B', 'Vino'), ('C', 'Tipo a efectos de IVA'),
            ('D', 'Precio de compra de la botella, sin IVA (€)'),
            ('E', 'Formato de la botella (cl)'),
            ('F', 'Servicio de copa (cl)'),
            ('G', 'Copas por botella (n.º)'),
            ('H', 'PVP de la botella en sala, sin IVA (€)'),
            ('I', 'PVP de la copa en sala, sin IVA (€)'),
            ('J', 'Botellas vendidas al mes (n.º)'),
            ('K', 'Copas vendidas al mes (n.º)'),
            ('L', 'Coste por copa (€)'),
            ('M', 'Margen por botella (€)'),
            ('N', 'Margen por copa (€)'),
            ('O', 'Beverage cost de la botella (%)'),
            ('P', 'Beverage cost de la copa (%)'),
            ('Q', 'Ventas del mes, sin IVA (€)'),
            ('R', 'Coste del mes (€)'),
            ('S', 'Margen de contribución del mes (€)'),
            ('T', 'Beverage cost del mes (%)'),
            ('U', 'IVA repercutido en sala (%)'),
            ('V', 'PVP de la botella con IVA en sala (€)'),
            ('W', 'PVP de la copa con IVA en sala (€)'),
            ('X', 'IVA repercutido para llevar (%)'),
            ('Y', 'PVP de la botella con IVA para llevar (€)')]
    cabecera(ws, V_CAB, cols, altura=62)
    anchos(ws, dict(zip('ABCDEFGHIJKLMNOPQRSTUVWXY',
                        (5, 34, 18, 15, 13, 13, 12, 15, 14, 13, 13, 12, 13,
                         12, 14, 13, 14, 13, 15, 13, 12, 15, 15, 13, 16))))

    for i in range(CAP_VINOS):
        r = V_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=FMT_ENT)
        if i < len(D.VINOS):
            (nombre, compra, formato, pvp_bot, pvp_copa, copas, uds_bot,
             uds_copa) = D.VINOS[i]
            servicio = round(formato / copas, 2)  # cl por copa que da ese nº
            motor.val(ws, 'B%d' % r, nombre, verde_=True)
            motor.val(ws, 'C%d' % r, TIPOS[2], verde_=True)
            motor.val(ws, 'D%d' % r, compra, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'E%d' % r, float(formato), fmt=FMT_CANT, verde_=True)
            motor.val(ws, 'F%d' % r, servicio, fmt=FMT_CANT, verde_=True)
            motor.val(ws, 'H%d' % r, pvp_bot, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'I%d' % r, pvp_copa, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'J%d' % r, uds_bot, fmt=FMT_ENT, verde_=True)
            motor.val(ws, 'K%d' % r, uds_copa, fmt=FMT_ENT, verde_=True)
        else:
            ws['D%d' % r].number_format = FMT_EUR
            ws['E%d' % r].number_format = FMT_CANT
            ws['F%d' % r].number_format = FMT_CANT
            ws['H%d' % r].number_format = FMT_EUR
            ws['I%d' % r].number_format = FMT_EUR
            ws['J%d' % r].number_format = FMT_ENT
            ws['K%d' % r].number_format = FMT_ENT
            motor.verde(ws, 'B%d:F%d' % (r, r))
            motor.verde(ws, 'H%d:K%d' % (r, r))
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF(OR($E{r}="",$F{r}="",$F{r}=0),"",$E{r}/$F{r}),"")'
                .format(r=r), fmt=FMT_CANT)
        motor.f(ws, 'L%d' % r,
                '=IFERROR(IF(OR($D{r}="",$G{r}="",$G{r}=0),"",$D{r}/$G{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'M%d' % r,
                '=IFERROR(IF(OR($H{r}="",$D{r}=""),"",$H{r}-$D{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'N%d' % r,
                '=IFERROR(IF(OR($I{r}="",$L{r}=""),"",$I{r}-$L{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'O%d' % r,
                '=IFERROR(IF(OR($D{r}="",$H{r}="",$H{r}=0),"",$D{r}/$H{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'P%d' % r,
                '=IFERROR(IF(OR($L{r}="",$I{r}="",$I{r}=0),"",$L{r}/$I{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'Q%d' % r,
                '=IFERROR(IF(OR($H{r}="",$I{r}="",$J{r}="",$K{r}=""),"",'
                '$H{r}*$J{r}+$I{r}*$K{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'R%d' % r,
                '=IFERROR(IF(OR($D{r}="",$L{r}="",$J{r}="",$K{r}=""),"",'
                '$D{r}*$J{r}+$L{r}*$K{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'S%d' % r,
                '=IFERROR(IF(OR($Q{r}="",$R{r}=""),"",$Q{r}-$R{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'T%d' % r,
                '=IFERROR(IF(OR($Q{r}="",$Q{r}=0,$R{r}=""),"",$R{r}/$Q{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'U%d' % r, iva_sala(r), fmt=FMT_PCT)
        motor.f(ws, 'V%d' % r,
                '=IFERROR(IF(OR($H{r}="",$U{r}=""),"",$H{r}*(1+$U{r})),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'W%d' % r,
                '=IFERROR(IF(OR($I{r}="",$U{r}=""),"",$I{r}*(1+$U{r})),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'X%d' % r, iva_llevar(r), fmt=FMT_PCT)
        motor.f(ws, 'Y%d' % r,
                '=IFERROR(IF(OR($H{r}="",$X{r}=""),"",$H{r}*(1+$X{r})),"")'
                .format(r=r), fmt=FMT_EUR)

    motor.val(ws, 'B%d' % V_TOT, 'TOTAL VINOS', bold=True)
    for col, fmt in (('J', FMT_ENT), ('K', FMT_ENT), ('Q', FMT_EUR),
                     ('R', FMT_EUR), ('S', FMT_EUR)):
        motor.f(ws, '%s%d' % (col, V_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=V_INI, b=V_FIN), fmt=fmt, bold=True)
    motor.f(ws, 'T%d' % V_TOT,
            '=IFERROR(IF(OR($Q${t}="",$Q${t}=0,$R${t}=""),"",$R${t}/$Q${t}),"")'
            .format(t=V_TOT), fmt=FMT_PCT, bold=True)
    fila_total(ws, V_TOT, 'A', 'Y')

    objetivo = objetivo_mirror(ws, V_OBJ, 'Parámetros!$B$17',
                               'Objetivo de beverage cost de los vinos (%)')
    semaforo_objetivo(ws, 'T{a}:T{b}'.format(a=V_INI, b=V_FIN), 'T', objetivo,
                      V_INI)
    semaforo_objetivo(ws, 'T{t}:T{t}'.format(t=V_TOT), 'T', objetivo, V_TOT)

    filas = list(range(V_INI, V_FIN + 1))
    motor.dv_lista(ws, ['C%d' % r for r in filas], list(TIPOS),
                   titulo='Tipo a efectos de IVA')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'DHI' for r in filas],
                      minimo=0, titulo='Importe (€)')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'EF' for r in filas],
                      minimo=0.1, maximo=10000, titulo='Centilitros')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'JK' for r in filas],
                      minimo=0, titulo='Unidades vendidas')
    motor.val(ws, 'B%d' % (V_OBJ + 2),
              'El servicio de copa es lo que sirves de verdad: cámbialo y '
              'cambian las copas por botella, el coste por copa y el beverage '
              'cost. Es la palanca más barata de toda la bodega.')
    ws.freeze_panes = 'D5'
    pagina(ws, titulos='$%d:$%d' % (V_CAB, V_CAB))
    return ws


# --------------------------------------------------------------------------
# Cervezas y Refrescos
# --------------------------------------------------------------------------
C_CAB, C_INI = 4, 5
CAP_CERVEZAS = 15         # capacidad de la hoja frente a las 7 sembradas
C_FIN = C_INI + CAP_CERVEZAS - 1
C_TOT = C_FIN + 1
C_OBJ = C_TOT + 2

#: Tipo a efectos de IVA y unidad de medida de cada fila del juego de datos.
#: El zumo se mide en ml (1 kg de naranjas → 1.000 ml de zumo, servicio de 250
#: ml); el resto, en cl. La cerveza SIN alcohol no es bebida alcohólica ni
#: lleva azúcares añadidos: para llevar va al tipo reducido.
CERVEZAS_META = [
    (TIPOS[2], 'cl'),      # barril, caña
    (TIPOS[2], 'cl'),      # barril, jarra
    (TIPOS[2], 'cl'),      # tercio
    (TIPOS[0], 'cl'),      # sin alcohol
    (TIPOS[1], 'cl'),      # refresco de cola
    (TIPOS[0], 'cl'),      # agua mineral
    (TIPOS[0], 'ml'),      # zumo de naranja natural
]


def hoja_cervezas(wb):
    ws = wb.create_sheet('Cervezas y Refrescos')
    encabezar(ws, 'Cervezas y refrescos',
              nota='El coste por servicio sale del precio de la unidad de '
                   'compra por la parte que sirves: un barril de 30 L da 120 '
                   'cañas de 25 cl.')
    cols = [('A', '#'), ('B', 'Referencia'), ('C', 'Formato de compra'),
            ('D', 'Tipo a efectos de IVA'), ('E', 'Unidad de medida'),
            ('F', 'Precio de compra de la unidad, sin IVA (€)'),
            ('G', 'Contenido de la unidad de compra'),
            ('H', 'Servicio'),
            ('I', 'PVP del servicio en sala, sin IVA (€)'),
            ('J', 'Servicios vendidos al mes (n.º)'),
            ('K', 'Servicios por unidad de compra (n.º)'),
            ('L', 'Coste por servicio (€)'),
            ('M', 'Margen por servicio (€)'),
            ('N', 'Beverage cost (%)'),
            ('O', 'Ventas del mes, sin IVA (€)'),
            ('P', 'Coste del mes (€)'),
            ('Q', 'Margen de contribución del mes (€)'),
            ('R', 'IVA repercutido en sala (%)'),
            ('S', 'PVP del servicio con IVA en sala (€)'),
            ('T', 'IVA repercutido para llevar (%)'),
            ('U', 'PVP del servicio con IVA para llevar (€)')]
    cabecera(ws, C_CAB, cols, altura=62)
    anchos(ws, dict(zip('ABCDEFGHIJKLMNOPQRSTU',
                        (5, 30, 16, 18, 14, 16, 17, 12, 15, 14, 16, 13, 13,
                         13, 14, 13, 15, 13, 16, 13, 17))))

    for i in range(CAP_CERVEZAS):
        r = C_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=FMT_ENT)
        if i < len(D.CERVEZAS_REFRESCOS):
            (nombre, formato_compra, precio, contenido, servicio, pvp,
             uds) = D.CERVEZAS_REFRESCOS[i]
            tipo, unidad = CERVEZAS_META[i]
            motor.val(ws, 'B%d' % r, nombre, verde_=True)
            motor.val(ws, 'C%d' % r, formato_compra, verde_=True)
            motor.val(ws, 'D%d' % r, tipo, verde_=True)
            motor.val(ws, 'E%d' % r, unidad, verde_=True)
            motor.val(ws, 'F%d' % r, precio, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'G%d' % r, float(contenido), fmt=FMT_CANT,
                      verde_=True)
            motor.val(ws, 'H%d' % r, float(servicio), fmt=FMT_CANT,
                      verde_=True)
            motor.val(ws, 'I%d' % r, pvp, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'J%d' % r, uds, fmt=FMT_ENT, verde_=True)
        else:
            ws['F%d' % r].number_format = FMT_EUR
            ws['G%d' % r].number_format = FMT_CANT
            ws['H%d' % r].number_format = FMT_CANT
            ws['I%d' % r].number_format = FMT_EUR
            ws['J%d' % r].number_format = FMT_ENT
            motor.verde(ws, 'B%d:J%d' % (r, r))
        motor.f(ws, 'K%d' % r,
                '=IFERROR(IF(OR($G{r}="",$H{r}="",$H{r}=0),"",$G{r}/$H{r}),"")'
                .format(r=r), fmt=FMT_CANT)
        motor.f(ws, 'L%d' % r,
                '=IFERROR(IF(OR($F{r}="",$G{r}="",$G{r}=0,$H{r}=""),"",'
                '$F{r}*$H{r}/$G{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'M%d' % r,
                '=IFERROR(IF(OR($I{r}="",$L{r}=""),"",$I{r}-$L{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'N%d' % r,
                '=IFERROR(IF(OR($L{r}="",$I{r}="",$I{r}=0),"",$L{r}/$I{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'O%d' % r,
                '=IFERROR(IF(OR($I{r}="",$J{r}=""),"",$I{r}*$J{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'P%d' % r,
                '=IFERROR(IF(OR($L{r}="",$J{r}=""),"",$L{r}*$J{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'Q%d' % r,
                '=IFERROR(IF(OR($O{r}="",$P{r}=""),"",$O{r}-$P{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'R%d' % r, iva_sala(r, 'D'), fmt=FMT_PCT)
        motor.f(ws, 'S%d' % r,
                '=IFERROR(IF(OR($I{r}="",$R{r}=""),"",$I{r}*(1+$R{r})),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'T%d' % r, iva_llevar(r, 'D'), fmt=FMT_PCT)
        motor.f(ws, 'U%d' % r,
                '=IFERROR(IF(OR($I{r}="",$T{r}=""),"",$I{r}*(1+$T{r})),"")'
                .format(r=r), fmt=FMT_EUR)

    motor.val(ws, 'B%d' % C_TOT, 'TOTAL CERVEZAS Y REFRESCOS', bold=True)
    for col, fmt in (('J', FMT_ENT), ('O', FMT_EUR), ('P', FMT_EUR),
                     ('Q', FMT_EUR)):
        motor.f(ws, '%s%d' % (col, C_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=C_INI, b=C_FIN), fmt=fmt, bold=True)
    motor.f(ws, 'N%d' % C_TOT,
            '=IFERROR(IF(OR($O${t}="",$O${t}=0,$P${t}=""),"",$P${t}/$O${t}),"")'
            .format(t=C_TOT), fmt=FMT_PCT, bold=True)
    fila_total(ws, C_TOT, 'A', 'U')

    objetivo = objetivo_mirror(ws, C_OBJ, 'Parámetros!$B$18',
                               'Objetivo de beverage cost de cervezas y refrescos (%)')
    semaforo_objetivo(ws, 'N{a}:N{b}'.format(a=C_INI, b=C_FIN), 'N', objetivo,
                      C_INI)
    semaforo_objetivo(ws, 'N{t}:N{t}'.format(t=C_TOT), 'N', objetivo, C_TOT)

    filas = list(range(C_INI, C_FIN + 1))
    motor.dv_lista(ws, ['D%d' % r for r in filas], list(TIPOS),
                   titulo='Tipo a efectos de IVA')
    motor.dv_lista(ws, ['E%d' % r for r in filas], ['cl', 'ml', 'g'],
                   titulo='Unidad de medida')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'FI' for r in filas],
                      minimo=0, titulo='Importe (€)')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'GH' for r in filas],
                      minimo=0.1, maximo=100000, titulo='Cantidad')
    motor.dv_numerica(ws, ['J%d' % r for r in filas], minimo=0,
                      titulo='Servicios vendidos')
    motor.val(ws, 'B%d' % (C_OBJ + 2),
              'Las dos primeras filas son el MISMO barril servido en caña de '
              '25 cl y en jarra de 50 cl: el coste por servicio se dobla, el '
              'precio no siempre. Míralo antes de poner la jarra en la carta.')
    motor.val(ws, 'B%d' % (C_OBJ + 3),
              'El zumo natural se mide en mililitros porque su unidad de '
              'compra son kilos de naranja: 1 kg da unos 1.000 ml de zumo y el '
              'servicio es de 250 ml. Ajusta el rendimiento con tu exprimidor.')
    ws.freeze_panes = 'D5'
    pagina(ws, titulos='$%d:$%d' % (C_CAB, C_CAB))
    return ws


# --------------------------------------------------------------------------
# Destilados y Cócteles
# --------------------------------------------------------------------------
DE_CAB, DE_INI = 4, 5
CAP_DESTILADOS = 12       # capacidad de la hoja frente a los 5 sembrados
DE_FIN = DE_INI + CAP_DESTILADOS - 1
DE_TOT = DE_FIN + 1
CI_TIT = DE_TOT + 2                    # título de la tabla de ingredientes
CI_CAB = CI_TIT + 1
CI_INI = CI_CAB + 1
LINEAS_POR_COCTEL = 4
CAP_COCTELES = 8          # capacidad de la hoja frente a los 4 sembrados
CI_FIN = CI_INI + CAP_COCTELES * LINEAS_POR_COCTEL - 1
CO_TIT = CI_FIN + 2
CO_CAB = CO_TIT + 1
CO_INI = CO_CAB + 1
CO_FIN = CO_INI + CAP_COCTELES - 1
CO_TOT = CO_FIN + 1
DE_OBJ = CO_TOT + 2


def hoja_destilados(wb):
    ws = wb.create_sheet('Destilados y Cócteles')
    encabezar(ws, 'Destilados y cócteles')
    anchos(ws, dict(zip('ABCDEFGHIJKLMNOPQRS',
                        (5, 30, 22, 16, 14, 15, 18, 16, 14, 15, 16, 15, 14,
                         14, 15, 14, 18, 13, 16))))

    seccion(ws, 'A%d' % (DE_CAB - 1), 'COMBINADOS: DESTILADO + MEZCLA')
    cabecera(ws, DE_CAB, [
        ('A', '#'), ('B', 'Destilado'), ('C', 'Tipo a efectos de IVA'),
        ('D', 'Precio de compra de la botella, sin IVA (€)'),
        ('E', 'Formato de la botella (cl)'), ('F', 'Servicio de copa (cl)'),
        ('G', 'Coste de la mezcla por copa (€)'),
        ('H', 'PVP del combinado en sala, sin IVA (€)'),
        ('I', 'Copas vendidas al mes (n.º)'),
        ('J', 'Copas por botella (n.º)'),
        ('K', 'Coste del destilado por copa (€)'),
        ('L', 'Coste total por copa (€)'), ('M', 'Margen por copa (€)'),
        ('N', 'Beverage cost (%)'), ('O', 'Ventas del mes, sin IVA (€)'),
        ('P', 'Coste del mes (€)'),
        ('Q', 'Margen de contribución del mes (€)'),
        ('R', 'IVA repercutido en sala (%)'),
        ('S', 'PVP del combinado con IVA en sala (€)')], altura=62)

    for i in range(CAP_DESTILADOS):
        r = DE_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=FMT_ENT)
        if i < len(D.DESTILADOS):
            nombre, compra, cl_copa, mezcla, pvp, uds = D.DESTILADOS[i]
            # El vermut de grifo se compra en botella de 1 L: lo confirma el
            # propio juego de datos (el negroni valora el vermut a 7,90 €/L y
            # la botella cuesta 7,90 €). El resto son botellas de 70 cl.
            formato = 100.0 if '1 L' in nombre else 70.0
            motor.val(ws, 'B%d' % r, nombre, verde_=True)
            motor.val(ws, 'C%d' % r, TIPOS[2], verde_=True)
            motor.val(ws, 'D%d' % r, compra, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'E%d' % r, formato, fmt=FMT_CANT, verde_=True)
            motor.val(ws, 'F%d' % r, float(cl_copa), fmt=FMT_CANT,
                      verde_=True)
            motor.val(ws, 'G%d' % r, mezcla, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'H%d' % r, pvp, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'I%d' % r, uds, fmt=FMT_ENT, verde_=True)
        else:
            ws['D%d' % r].number_format = FMT_EUR
            ws['E%d' % r].number_format = FMT_CANT
            ws['F%d' % r].number_format = FMT_CANT
            ws['G%d' % r].number_format = FMT_EUR
            ws['H%d' % r].number_format = FMT_EUR
            ws['I%d' % r].number_format = FMT_ENT
            motor.verde(ws, 'B%d:I%d' % (r, r))
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF(OR($E{r}="",$F{r}="",$F{r}=0),"",$E{r}/$F{r}),"")'
                .format(r=r), fmt=FMT_CANT)
        motor.f(ws, 'K%d' % r,
                '=IFERROR(IF(OR($D{r}="",$E{r}="",$E{r}=0,$F{r}=""),"",'
                '$D{r}*$F{r}/$E{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'L%d' % r,
                '=IFERROR(IF($K{r}="","",$K{r}+IF($G{r}="",0,$G{r})),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'M%d' % r,
                '=IFERROR(IF(OR($H{r}="",$L{r}=""),"",$H{r}-$L{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'N%d' % r,
                '=IFERROR(IF(OR($L{r}="",$H{r}="",$H{r}=0),"",$L{r}/$H{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'O%d' % r,
                '=IFERROR(IF(OR($H{r}="",$I{r}=""),"",$H{r}*$I{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'P%d' % r,
                '=IFERROR(IF(OR($L{r}="",$I{r}=""),"",$L{r}*$I{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'Q%d' % r,
                '=IFERROR(IF(OR($O{r}="",$P{r}=""),"",$O{r}-$P{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'R%d' % r, iva_sala(r), fmt=FMT_PCT)
        motor.f(ws, 'S%d' % r,
                '=IFERROR(IF(OR($H{r}="",$R{r}=""),"",$H{r}*(1+$R{r})),"")'
                .format(r=r), fmt=FMT_EUR)

    motor.val(ws, 'B%d' % DE_TOT, 'TOTAL COMBINADOS', bold=True)
    for col, fmt in (('I', FMT_ENT), ('O', FMT_EUR), ('P', FMT_EUR),
                     ('Q', FMT_EUR)):
        motor.f(ws, '%s%d' % (col, DE_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=DE_INI, b=DE_FIN), fmt=fmt, bold=True)
    motor.f(ws, 'N%d' % DE_TOT,
            '=IFERROR(IF(OR($O${t}="",$O${t}=0,$P${t}=""),"",$P${t}/$O${t}),"")'
            .format(t=DE_TOT), fmt=FMT_PCT, bold=True)
    fila_total(ws, DE_TOT, 'A', 'S')

    # --- cócteles: líneas de ingredientes ---
    seccion(ws, 'A%d' % CI_TIT, 'CÓCTELES: COSTE INGREDIENTE A INGREDIENTE')
    cabecera(ws, CI_CAB, [('A', '#'), ('B', 'Cóctel'), ('C', 'Ingrediente'),
                          ('D', 'Cantidad (cl)'),
                          ('E', 'Precio del ingrediente (€/L)'),
                          ('F', 'Coste de la línea (€)')], altura=44)
    for i in range(CAP_COCTELES):
        nombre, lineas = (D.COCTELES[i][0], D.COCTELES[i][1]) \
            if i < len(D.COCTELES) else (None, [])
        for j in range(LINEAS_POR_COCTEL):
            r = CI_INI + i * LINEAS_POR_COCTEL + j
            motor.val(ws, 'A%d' % r, j + 1, fmt=FMT_ENT)
            if nombre is not None:
                motor.val(ws, 'B%d' % r, nombre, verde_=True)
            else:
                motor.verde(ws, 'B%d' % r)
            if j < len(lineas):
                ingrediente, cantidad, precio_litro = lineas[j]
                motor.val(ws, 'C%d' % r, ingrediente, verde_=True)
                motor.val(ws, 'D%d' % r, float(cantidad), fmt=FMT_CANT,
                          verde_=True)
                motor.val(ws, 'E%d' % r, precio_litro, fmt=FMT_EUR,
                          verde_=True)
            else:
                motor.verde(ws, 'C%d' % r)
                motor.verde(ws, 'D%d' % r)
                ws['D%d' % r].number_format = FMT_CANT
                motor.verde(ws, 'E%d' % r)
                ws['E%d' % r].number_format = FMT_EUR
            motor.f(ws, 'F%d' % r,
                    '=IFERROR(IF(OR($D{r}="",$E{r}=""),"",$D{r}*$E{r}/100),"")'
                    .format(r=r), fmt=FMT_EUR)
    motor.val(ws, 'B%d' % (CI_FIN + 1),
              'La cantidad va en centilitros y el precio en euros por litro, '
              'así que el coste de la línea es cantidad × precio ÷ 100.')

    # --- cócteles: resumen por cóctel ---
    seccion(ws, 'A%d' % CO_TIT, 'CÓCTELES: RESUMEN POR CÓCTEL')
    cabecera(ws, CO_CAB, [
        ('A', '#'), ('B', 'Cóctel'), ('C', 'Tipo a efectos de IVA'),
        ('D', 'Coste por copa (€)'), ('E', 'PVP en sala, sin IVA (€)'),
        ('F', 'Copas vendidas al mes (n.º)'),
        ('G', '¿Suma al total de la bodega?'),
        ('H', 'Margen por copa (€)'), ('I', 'Beverage cost (%)'),
        ('J', 'Ventas del mes, sin IVA (€)'), ('K', 'Coste del mes (€)'),
        ('L', 'Margen de contribución del mes (€)'),
        ('M', 'IVA repercutido en sala (%)'),
        ('N', 'PVP con IVA en sala (€)')], altura=62)
    for i in range(CAP_COCTELES):
        r = CO_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=FMT_ENT)
        if i < len(D.COCTELES):
            nombre, lineas, pvp, uds = D.COCTELES[i]
            # El gin tonic ya está contado arriba como combinado (misma copa,
            # mismo precio, mismas unidades): aquí se desglosa para comparar
            # los dos métodos de costeo, pero NO vuelve a sumar al total de
            # la bodega.
            suma = 'No' if 'Gin tonic' in nombre else 'Sí'
            motor.val(ws, 'B%d' % r, nombre, verde_=True)
            motor.val(ws, 'C%d' % r, TIPOS[2], verde_=True)
            motor.val(ws, 'E%d' % r, pvp, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'F%d' % r, uds, fmt=FMT_ENT, verde_=True)
            motor.val(ws, 'G%d' % r, suma, verde_=True)
        else:
            ws['E%d' % r].number_format = FMT_EUR
            ws['F%d' % r].number_format = FMT_ENT
            motor.verde(ws, 'B%d:C%d' % (r, r))
            motor.verde(ws, 'E%d:G%d' % (r, r))
        # SUMPRODUCT, no SUMIF/COUNTIF: el criterio es $B{r}, que en las filas
        # libres llega vacío, y COUNTIF con criterio vacío revienta en pycel
        # («Couldn't parse criteria: None») dejando el libro sin cache.
        motor.f(ws, 'D%d' % r,
                '=IFERROR(IF(OR($B{r}="",SUMPRODUCT(--($B${a}:$B${b}=$B{r}))'
                '=0),"",SUMPRODUCT(--($B${a}:$B${b}=$B{r}),$F${a}:$F${b})),'
                '"")'.format(r=r, a=CI_INI, b=CI_FIN), fmt=FMT_EUR)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF(OR($E{r}="",$D{r}=""),"",$E{r}-$D{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF(OR($D{r}="",$E{r}="",$E{r}=0),"",$D{r}/$E{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF(OR($E{r}="",$F{r}=""),"",$E{r}*$F{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'K%d' % r,
                '=IFERROR(IF(OR($D{r}="",$F{r}=""),"",$D{r}*$F{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'L%d' % r,
                '=IFERROR(IF(OR($J{r}="",$K{r}=""),"",$J{r}-$K{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'M%d' % r, iva_sala(r), fmt=FMT_PCT)
        motor.f(ws, 'N%d' % r,
                '=IFERROR(IF(OR($E{r}="",$M{r}=""),"",$E{r}*(1+$M{r})),"")'
                .format(r=r), fmt=FMT_EUR)

    motor.val(ws, 'B%d' % CO_TOT, 'TOTAL CÓCTELES QUE SUMAN A LA BODEGA',
              bold=True)
    for col, fmt in (('J', FMT_EUR), ('K', FMT_EUR), ('L', FMT_EUR)):
        motor.f(ws, '%s%d' % (col, CO_TOT),
                '=IFERROR(IF(COUNTIF($G${a}:$G${b},"Sí")=0,"",'
                'SUMIF($G${a}:$G${b},"Sí",${c}${a}:${c}${b})),"")'
                .format(a=CO_INI, b=CO_FIN, c=col), fmt=fmt, bold=True)
    motor.f(ws, 'I%d' % CO_TOT,
            '=IFERROR(IF(OR($J${t}="",$J${t}=0,$K${t}=""),"",$K${t}/$J${t}),"")'
            .format(t=CO_TOT), fmt=FMT_PCT, bold=True)
    fila_total(ws, CO_TOT, 'A', 'N')

    objetivo = objetivo_mirror(ws, DE_OBJ, 'Parámetros!$B$19',
                               'Objetivo de beverage cost de destilados y cócteles (%)')
    semaforo_objetivo(ws, 'N{a}:N{b}'.format(a=DE_INI, b=DE_FIN), 'N',
                      objetivo, DE_INI)
    semaforo_objetivo(ws, 'N{t}:N{t}'.format(t=DE_TOT), 'N', objetivo, DE_TOT)
    semaforo_objetivo(ws, 'I{a}:I{b}'.format(a=CO_INI, b=CO_FIN), 'I',
                      objetivo, CO_INI)
    semaforo_objetivo(ws, 'I{t}:I{t}'.format(t=CO_TOT), 'I', objetivo, CO_TOT)

    motor.val(ws, 'B%d' % (DE_OBJ + 2),
              'El gin tonic aparece en las dos tablas a propósito: arriba '
              'costeado como destilado + mezcla y abajo desglosado ingrediente '
              'a ingrediente. Compara los dos costes; para que no se cuente '
              'dos veces, abajo está marcado como «No» en «¿Suma al total de '
              'la bodega?».')
    motor.val(ws, 'B%d' % (DE_OBJ + 3),
              'Si un vino se usa como ingrediente de un cóctel (el cava del '
              'spritz), cuéntalo aquí y deja en la hoja «Vinos» sólo lo que '
              'vendas COMO vino: si no, la misma botella se paga dos veces.')

    filas_de = list(range(DE_INI, DE_FIN + 1))
    filas_co = list(range(CO_INI, CO_FIN + 1))
    motor.dv_lista(ws, ['C%d' % r for r in filas_de + filas_co], list(TIPOS),
                   titulo='Tipo a efectos de IVA')
    motor.dv_lista(ws, ['G%d' % r for r in filas_co], ['Sí', 'No'],
                   titulo='¿Suma al total?')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'DGH' for r in filas_de]
                      + ['E%d' % r for r in filas_co]
                      + ['E%d' % r for r in range(CI_INI, CI_FIN + 1)],
                      minimo=0, titulo='Importe (€)')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'EF' for r in filas_de]
                      + ['D%d' % r for r in range(CI_INI, CI_FIN + 1)],
                      minimo=0.1, maximo=10000, titulo='Centilitros')
    motor.dv_numerica(ws, ['I%d' % r for r in filas_de]
                      + ['F%d' % r for r in filas_co],
                      minimo=0, titulo='Copas vendidas')
    ws.freeze_panes = 'D5'
    pagina(ws, titulos='$%d:$%d' % (DE_CAB, DE_CAB))
    return ws


# --------------------------------------------------------------------------
# Resumen Bodega
# --------------------------------------------------------------------------
R_CAB, R_INI = 4, 5
R_FIN = 7
R_TOT = 8


def hoja_resumen(wb):
    ws = wb.create_sheet('Resumen Bodega')
    anchos(ws, {'A': 26, 'B': 18, 'C': 16, 'D': 18, 'E': 18, 'F': 20,
                'G': 20, 'H': 22})
    encabezar(ws, 'Resumen de la bodega',
              nota='Todo se calcula desde las tres hojas de producto: aquí no '
                   'se teclea nada.')
    cabecera(ws, R_CAB, [
        ('A', 'Categoría'), ('B', 'Ventas del mes, sin IVA (€)'),
        ('C', 'Coste del mes (€)'), ('D', 'Beverage cost ponderado (%)'),
        ('E', 'Objetivo de beverage cost (%)'),
        ('F', 'Margen de contribución del mes (€)'),
        ('G', 'Peso sobre las ventas de bodega (%)'),
        ('H', 'Margen que ganarías si llegaras al objetivo (€/mes)')],
             altura=62)

    ventas = ('=IF(Vinos!$Q${0}="","",Vinos!$Q${0})'.format(V_TOT),
              "=IF('Cervezas y Refrescos'!$O${0}=\"\",\"\","
              "'Cervezas y Refrescos'!$O${0})".format(C_TOT),
              "=IFERROR(IF(AND('Destilados y Cócteles'!$O${0}=\"\","
              "'Destilados y Cócteles'!$J${1}=\"\"),\"\","
              "IF('Destilados y Cócteles'!$O${0}=\"\",0,"
              "'Destilados y Cócteles'!$O${0})"
              "+IF('Destilados y Cócteles'!$J${1}=\"\",0,"
              "'Destilados y Cócteles'!$J${1})),\"\")".format(DE_TOT, CO_TOT))
    costes = ('=IF(Vinos!$R${0}="","",Vinos!$R${0})'.format(V_TOT),
              "=IF('Cervezas y Refrescos'!$P${0}=\"\",\"\","
              "'Cervezas y Refrescos'!$P${0})".format(C_TOT),
              "=IFERROR(IF(AND('Destilados y Cócteles'!$P${0}=\"\","
              "'Destilados y Cócteles'!$K${1}=\"\"),\"\","
              "IF('Destilados y Cócteles'!$P${0}=\"\",0,"
              "'Destilados y Cócteles'!$P${0})"
              "+IF('Destilados y Cócteles'!$K${1}=\"\",0,"
              "'Destilados y Cócteles'!$K${1})),\"\")".format(DE_TOT, CO_TOT))
    objetivos = ('=IF(Parámetros!$B$17="","",Parámetros!$B$17)',
                 '=IF(Parámetros!$B$18="","",Parámetros!$B$18)',
                 '=IF(Parámetros!$B$19="","",Parámetros!$B$19)')
    nombres = ('=Parámetros!$A$17', '=Parámetros!$A$18', '=Parámetros!$A$19')

    for i in range(3):
        r = R_INI + i
        motor.f(ws, 'A%d' % r, nombres[i], bold=True)
        motor.f(ws, 'B%d' % r, ventas[i], fmt=FMT_EUR)
        motor.f(ws, 'C%d' % r, costes[i], fmt=FMT_EUR)
        motor.f(ws, 'D%d' % r,
                '=IFERROR(IF(OR($B{r}="",$B{r}=0,$C{r}=""),"",$C{r}/$B{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'E%d' % r, objetivos[i], fmt=FMT_PCT)
        motor.f(ws, 'F%d' % r,
                '=IFERROR(IF(OR($B{r}="",$C{r}=""),"",$B{r}-$C{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF(OR($B{r}="",$B${t}="",$B${t}=0),"",'
                '$B{r}/$B${t}),"")'.format(r=r, t=R_TOT), fmt=FMT_PCT)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF(OR($B{r}="",$C{r}="",$E{r}=""),"",'
                'MAX(0,$C{r}-$B{r}*$E{r})),"")'.format(r=r), fmt=FMT_EUR)

    motor.val(ws, 'A%d' % R_TOT, 'TOTAL BODEGA', bold=True)
    for col in ('B', 'C', 'F', 'H'):
        motor.f(ws, '%s%d' % (col, R_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=R_INI, b=R_FIN), fmt=FMT_EUR, bold=True)
    motor.f(ws, 'D%d' % R_TOT,
            '=IFERROR(IF(OR($B${t}="",$B${t}=0,$C${t}=""),"",$C${t}/$B${t}),"")'
            .format(t=R_TOT), fmt=FMT_PCT, bold=True)
    motor.f(ws, 'G%d' % R_TOT,
            '=IF($B${t}="","",1)'.format(t=R_TOT), fmt=FMT_PCT, bold=True)
    fila_total(ws, R_TOT, 'A', 'H')

    cf_expresion(ws, 'D%d:D%d' % (R_INI, R_TOT),
                 '=AND(ISNUMBER($D%d),ISNUMBER($E%d),$D%d>$E%d)'
                 % (R_INI, R_INI, R_INI, R_INI),
                 motor.CF_ROJO_BG, motor.CF_ROJO_FG)
    cf_expresion(ws, 'D%d:D%d' % (R_INI, R_TOT),
                 '=AND(ISNUMBER($D%d),ISNUMBER($E%d),$D%d<=$E%d)'
                 % (R_INI, R_INI, R_INI, R_INI),
                 motor.CF_VERDE_BG, motor.CF_VERDE_FG)

    seccion(ws, 'A10', 'CÓMO SE LEE')
    lecturas = [
        'La última columna es la traducción a euros del semáforo: lo que dejas '
        'de ganar cada mes por estar por encima del objetivo en esa categoría. '
        'Si estás en objetivo o por debajo, es cero.',
        'La fila TOTAL BODEGA no tiene objetivo propio a propósito: la media de '
        'la bodega depende de tu mix de ventas, no de tu política de precios. '
        'Se gestiona categoría a categoría.',
        'El margen de esta hoja es margen de CONTRIBUCIÓN de la bebida: no '
        'descuenta personal de barra ni estructura. Eso lo mide el libro '
        '«cuadro-de-mando-prime-cost.xlsx».',
        'Referencia española: la bebida pesa alrededor del 34,5 % de coste '
        'sobre los ingresos de bebida frente al 28 % de la comida '
        '(CaixaBankLab × elBulliFoundation). Por eso la bodega se lleva como '
        'una cuenta de resultados propia y no como un apéndice de la cocina.',
    ]
    fila = 11
    for texto in lecturas:
        motor.val(ws, 'A' + str(fila), texto, wrap=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
    pagina(ws)
    return ws


# --------------------------------------------------------------------------
def mapa():
    m = {'fichero': NOMBRE + '.xlsx', 'hojas': {}}
    m['hojas']['Parámetros'] = {
        'celdas': {
            'IVA repercutido en sala, comida': 'B6',
            'IVA repercutido en sala, refresco azucarado': 'C6',
            'IVA repercutido en sala, bebida alcohólica': 'D6',
            'IVA repercutido para llevar, comida': 'B7',
            'IVA repercutido para llevar, refresco azucarado': 'C7',
            'IVA repercutido para llevar, bebida alcohólica': 'D7',
            'IVA repercutido en delivery, comida': 'B8',
            'IVA repercutido en delivery, refresco azucarado': 'C8',
            'IVA repercutido en delivery, bebida alcohólica': 'D8',
            'Objetivo de beverage cost de los vinos': 'B17',
            'Objetivo de beverage cost de cervezas y refrescos': 'B18',
            'Objetivo de beverage cost de destilados y cócteles': 'B19',
        },
        'tablas': [
            {'titulo': 'Matriz de IVA repercutido por canal y tipo de producto',
             'cols': [['Canal', 'A', 'txt'], ['Comida', 'B', 'pct1'],
                      ['Refresco/azucarada', 'C', 'pct1'],
                      ['Bebida alcohólica', 'D', 'pct1']],
             'filas': [6, 8]},
            {'titulo': 'Objetivo de beverage cost por categoría',
             'cols': [['Categoría', 'A', 'txt'], ['Objetivo (%)', 'B', 'pct1']],
             'filas': [17, 19]},
        ],
    }

    celdas_v = {
        'Botellas vendidas al mes, total': 'J%d' % V_TOT,
        'Copas vendidas al mes, total': 'K%d' % V_TOT,
        'Ventas de vino del mes, sin IVA': 'Q%d' % V_TOT,
        'Coste de vino del mes': 'R%d' % V_TOT,
        'Margen de contribución del vino del mes': 'S%d' % V_TOT,
        'Beverage cost ponderado de los vinos': 'T%d' % V_TOT,
        'Objetivo de beverage cost de los vinos (espejo)': 'E%d' % V_OBJ,
    }
    for i, v in enumerate(D.VINOS):
        r = V_INI + i
        celdas_v['Coste por copa de %s' % v[0]] = 'L%d' % r
        celdas_v['Margen por botella de %s' % v[0]] = 'M%d' % r
        celdas_v['Margen por copa de %s' % v[0]] = 'N%d' % r
        celdas_v['Beverage cost de la botella de %s' % v[0]] = 'O%d' % r
        celdas_v['Beverage cost de la copa de %s' % v[0]] = 'P%d' % r
        celdas_v['Beverage cost del mes de %s' % v[0]] = 'T%d' % r
        celdas_v['PVP con IVA en sala de la botella de %s' % v[0]] = 'V%d' % r
        celdas_v['PVP con IVA en sala de la copa de %s' % v[0]] = 'W%d' % r
        celdas_v['PVP con IVA para llevar de la botella de %s' % v[0]] = \
            'Y%d' % r
    m['hojas']['Vinos'] = {
        'celdas': celdas_v,
        'tablas': [{'titulo': 'Vinos: botella y copa',
                    'cols': [['#', 'A', 'num'], ['Vino', 'B', 'txt'],
                             ['Tipo a efectos de IVA', 'C', 'txt'],
                             ['Precio de compra de la botella, sin IVA (€)',
                              'D', 'eur'],
                             ['Formato de la botella (cl)', 'E', 'num1'],
                             ['Servicio de copa (cl)', 'F', 'num1'],
                             ['Copas por botella (n.º)', 'G', 'num1'],
                             ['PVP de la botella en sala, sin IVA (€)',
                              'H', 'eur'],
                             ['PVP de la copa en sala, sin IVA (€)',
                              'I', 'eur'],
                             ['Botellas vendidas al mes (n.º)', 'J', 'num'],
                             ['Copas vendidas al mes (n.º)', 'K', 'num'],
                             ['Coste por copa (€)', 'L', 'eur'],
                             ['Margen por botella (€)', 'M', 'eur'],
                             ['Margen por copa (€)', 'N', 'eur'],
                             ['Beverage cost de la botella (%)', 'O', 'pct1'],
                             ['Beverage cost de la copa (%)', 'P', 'pct1'],
                             ['Ventas del mes, sin IVA (€)', 'Q', 'eur'],
                             ['Coste del mes (€)', 'R', 'eur'],
                             ['Margen de contribución del mes (€)', 'S', 'eur'],
                             ['Beverage cost del mes (%)', 'T', 'pct1'],
                             ['IVA repercutido en sala (%)', 'U', 'pct1'],
                             ['PVP de la botella con IVA en sala (€)',
                              'V', 'eur'],
                             ['PVP de la copa con IVA en sala (€)', 'W', 'eur'],
                             ['IVA repercutido para llevar (%)', 'X', 'pct1'],
                             ['PVP de la botella con IVA para llevar (€)',
                              'Y', 'eur']],
                    'filas': [V_INI, V_FIN]}],
    }

    celdas_c = {
        'Ventas de cervezas y refrescos del mes, sin IVA': 'O%d' % C_TOT,
        'Coste de cervezas y refrescos del mes': 'P%d' % C_TOT,
        'Margen de contribución de cervezas y refrescos': 'Q%d' % C_TOT,
        'Beverage cost ponderado de cervezas y refrescos': 'N%d' % C_TOT,
        'Servicios vendidos al mes, total': 'J%d' % C_TOT,
        'Objetivo de beverage cost de cervezas y refrescos (espejo)':
            'E%d' % C_OBJ,
    }
    for i, cr in enumerate(D.CERVEZAS_REFRESCOS):
        r = C_INI + i
        eti = '%s (%s cl)' % (cr[0], cr[4])
        celdas_c['Servicios por unidad de compra de %s' % eti] = 'K%d' % r
        celdas_c['Coste por servicio de %s' % eti] = 'L%d' % r
        celdas_c['Margen por servicio de %s' % eti] = 'M%d' % r
        celdas_c['Beverage cost de %s' % eti] = 'N%d' % r
        celdas_c['Ventas del mes de %s' % eti] = 'O%d' % r
        celdas_c['PVP con IVA en sala de %s' % eti] = 'S%d' % r
        celdas_c['PVP con IVA para llevar de %s' % eti] = 'U%d' % r
    m['hojas']['Cervezas y Refrescos'] = {
        'celdas': celdas_c,
        'tablas': [{'titulo': 'Cervezas y refrescos',
                    'cols': [['#', 'A', 'num'], ['Referencia', 'B', 'txt'],
                             ['Formato de compra', 'C', 'txt'],
                             ['Tipo a efectos de IVA', 'D', 'txt'],
                             ['Unidad de medida', 'E', 'txt'],
                             ['Precio de compra de la unidad, sin IVA (€)',
                              'F', 'eur'],
                             ['Contenido de la unidad de compra', 'G', 'num1'],
                             ['Servicio', 'H', 'num1'],
                             ['PVP del servicio en sala, sin IVA (€)',
                              'I', 'eur'],
                             ['Servicios vendidos al mes (n.º)', 'J', 'num'],
                             ['Servicios por unidad de compra (n.º)',
                              'K', 'num1'],
                             ['Coste por servicio (€)', 'L', 'eur'],
                             ['Margen por servicio (€)', 'M', 'eur'],
                             ['Beverage cost (%)', 'N', 'pct1'],
                             ['Ventas del mes, sin IVA (€)', 'O', 'eur'],
                             ['Coste del mes (€)', 'P', 'eur'],
                             ['Margen de contribución del mes (€)', 'Q', 'eur'],
                             ['IVA repercutido en sala (%)', 'R', 'pct1'],
                             ['PVP del servicio con IVA en sala (€)',
                              'S', 'eur'],
                             ['IVA repercutido para llevar (%)', 'T', 'pct1'],
                             ['PVP del servicio con IVA para llevar (€)',
                              'U', 'eur']],
                    'filas': [C_INI, C_FIN]}],
    }

    celdas_d = {
        'Ventas de combinados del mes, sin IVA': 'O%d' % DE_TOT,
        'Coste de combinados del mes': 'P%d' % DE_TOT,
        'Margen de contribución de los combinados': 'Q%d' % DE_TOT,
        'Beverage cost ponderado de los combinados': 'N%d' % DE_TOT,
        'Ventas de cócteles que suman a la bodega': 'J%d' % CO_TOT,
        'Coste de cócteles que suman a la bodega': 'K%d' % CO_TOT,
        'Margen de contribución de los cócteles': 'L%d' % CO_TOT,
        'Beverage cost ponderado de los cócteles': 'I%d' % CO_TOT,
        'Objetivo de beverage cost de destilados y cócteles (espejo)':
            'E%d' % DE_OBJ,
    }
    for i, de in enumerate(D.DESTILADOS):
        r = DE_INI + i
        celdas_d['Copas por botella de %s' % de[0]] = 'J%d' % r
        celdas_d['Coste del destilado por copa de %s' % de[0]] = 'K%d' % r
        celdas_d['Coste total por copa de %s' % de[0]] = 'L%d' % r
        celdas_d['Margen por copa de %s' % de[0]] = 'M%d' % r
        celdas_d['Beverage cost de %s' % de[0]] = 'N%d' % r
        celdas_d['PVP con IVA en sala de %s' % de[0]] = 'S%d' % r
    for i, co in enumerate(D.COCTELES):
        r = CO_INI + i
        celdas_d['Coste por copa de %s' % co[0]] = 'D%d' % r
        celdas_d['Margen por copa de %s' % co[0]] = 'H%d' % r
        celdas_d['Beverage cost de %s' % co[0]] = 'I%d' % r
        celdas_d['PVP con IVA en sala de %s' % co[0]] = 'N%d' % r
    m['hojas']['Destilados y Cócteles'] = {
        'celdas': celdas_d,
        'tablas': [
            {'titulo': 'Combinados: destilado + mezcla',
             'cols': [['#', 'A', 'num'], ['Destilado', 'B', 'txt'],
                      ['Tipo a efectos de IVA', 'C', 'txt'],
                      ['Precio de compra de la botella, sin IVA (€)',
                       'D', 'eur'],
                      ['Formato de la botella (cl)', 'E', 'num1'],
                      ['Servicio de copa (cl)', 'F', 'num1'],
                      ['Coste de la mezcla por copa (€)', 'G', 'eur'],
                      ['PVP del combinado en sala, sin IVA (€)', 'H', 'eur'],
                      ['Copas vendidas al mes (n.º)', 'I', 'num'],
                      ['Copas por botella (n.º)', 'J', 'num1'],
                      ['Coste del destilado por copa (€)', 'K', 'eur'],
                      ['Coste total por copa (€)', 'L', 'eur'],
                      ['Margen por copa (€)', 'M', 'eur'],
                      ['Beverage cost (%)', 'N', 'pct1'],
                      ['Ventas del mes, sin IVA (€)', 'O', 'eur'],
                      ['Coste del mes (€)', 'P', 'eur'],
                      ['Margen de contribución del mes (€)', 'Q', 'eur'],
                      ['IVA repercutido en sala (%)', 'R', 'pct1'],
                      ['PVP del combinado con IVA en sala (€)', 'S', 'eur']],
             'filas': [DE_INI, DE_FIN]},
            {'titulo': 'Cócteles: coste ingrediente a ingrediente',
             'cols': [['#', 'A', 'num'], ['Cóctel', 'B', 'txt'],
                      ['Ingrediente', 'C', 'txt'],
                      ['Cantidad (cl)', 'D', 'num1'],
                      ['Precio del ingrediente (€/L)', 'E', 'eur'],
                      ['Coste de la línea (€)', 'F', 'eur']],
             'filas': [CI_INI, CI_FIN]},
            {'titulo': 'Cócteles: resumen por cóctel',
             'cols': [['#', 'A', 'num'], ['Cóctel', 'B', 'txt'],
                      ['Tipo a efectos de IVA', 'C', 'txt'],
                      ['Coste por copa (€)', 'D', 'eur'],
                      ['PVP en sala, sin IVA (€)', 'E', 'eur'],
                      ['Copas vendidas al mes (n.º)', 'F', 'num'],
                      ['¿Suma al total de la bodega?', 'G', 'txt'],
                      ['Margen por copa (€)', 'H', 'eur'],
                      ['Beverage cost (%)', 'I', 'pct1'],
                      ['Ventas del mes, sin IVA (€)', 'J', 'eur'],
                      ['Coste del mes (€)', 'K', 'eur'],
                      ['Margen de contribución del mes (€)', 'L', 'eur'],
                      ['IVA repercutido en sala (%)', 'M', 'pct1'],
                      ['PVP con IVA en sala (€)', 'N', 'eur']],
             'filas': [CO_INI, CO_FIN]},
        ],
    }

    m['hojas']['Resumen Bodega'] = {
        'celdas': {
            'Ventas de vinos del mes': 'B5',
            'Ventas de cervezas y refrescos del mes': 'B6',
            'Ventas de destilados y cócteles del mes': 'B7',
            'Ventas totales de bodega del mes': 'B8',
            'Coste de vinos del mes': 'C5',
            'Coste de cervezas y refrescos del mes': 'C6',
            'Coste de destilados y cócteles del mes': 'C7',
            'Coste total de bodega del mes': 'C8',
            'Beverage cost ponderado de vinos': 'D5',
            'Beverage cost ponderado de cervezas y refrescos': 'D6',
            'Beverage cost ponderado de destilados y cócteles': 'D7',
            'Beverage cost ponderado de toda la bodega': 'D8',
            'Objetivo de beverage cost de vinos': 'E5',
            'Objetivo de beverage cost de cervezas y refrescos': 'E6',
            'Objetivo de beverage cost de destilados y cócteles': 'E7',
            'Margen de contribución de vinos': 'F5',
            'Margen de contribución de cervezas y refrescos': 'F6',
            'Margen de contribución de destilados y cócteles': 'F7',
            'Margen de contribución total de la bodega': 'F8',
            'Peso de los vinos sobre las ventas de bodega': 'G5',
            'Peso de cervezas y refrescos sobre las ventas de bodega': 'G6',
            'Peso de destilados y cócteles sobre las ventas de bodega': 'G7',
            'Margen a recuperar si los vinos llegaran al objetivo': 'H5',
            'Margen a recuperar si cervezas y refrescos llegaran al objetivo':
                'H6',
            'Margen a recuperar si destilados y cócteles llegaran al objetivo':
                'H7',
            'Margen total a recuperar en la bodega': 'H8',
        },
        'tablas': [{'titulo': 'Resumen de la bodega por categoría',
                    'cols': [['Categoría', 'A', 'txt'],
                             ['Ventas del mes, sin IVA (€)', 'B', 'eur'],
                             ['Coste del mes (€)', 'C', 'eur'],
                             ['Beverage cost ponderado (%)', 'D', 'pct1'],
                             ['Objetivo de beverage cost (%)', 'E', 'pct1'],
                             ['Margen de contribución del mes (€)', 'F', 'eur'],
                             ['Peso sobre las ventas de bodega (%)',
                              'G', 'pct1'],
                             ['Margen que ganarías si llegaras al objetivo '
                              '(€/mes)', 'H', 'eur']],
                    'filas': [R_INI, R_TOT]}],
    }
    return m


def main():
    import json
    wb = Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_parametros(wb)
    hoja_vinos(wb)
    hoja_cervezas(wb)
    hoja_destilados(wb)
    hoja_resumen(wb)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · Versión 1.0 · septiembre 2026'

    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    destino = os.path.join(AQUI, 'build')
    os.makedirs(destino, exist_ok=True)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')
    wb.save(ruta)
    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)
    print('escrito:', ruta)
    print('fórmulas registradas:', len(motor.REGISTRO))


if __name__ == '__main__':
    main()
