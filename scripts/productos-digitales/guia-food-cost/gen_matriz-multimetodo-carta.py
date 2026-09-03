#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_matriz-multimetodo-carta.py — Libro 4 del pack «Guía Food Cost + Ingeniería
de Menú» (SPEC §2.2 fila 4 · decisión D6 · decisión D14 para el menú de precio
fijo).

Genera `build/matriz-multimetodo-carta.xlsx`:

  Instrucciones · Datos · Kasavana-Smith · Miller · Pavesic · Goal Value ·
  Comparativa · Menú Precio Fijo

Los cuatro métodos leen LA MISMA tabla de «Datos» y clasifican cada plato DENTRO
de su familia. «Comparativa» pone las cuatro lecturas lado a lado, cuenta
cuántas se salen de la mejor categoría y escribe un diagnóstico: cada método
mide dos de las tres variables (popularidad, margen en euros y food cost %), así
que discrepar es lo normal y es información, no ruido.

DECISIONES TÉCNICAS
-------------------
* **Nada de `SUMIF`/`COUNTIF` con la celda de familia como criterio.** pycel
  evalúa los dos brazos de un `IF` de forma ansiosa, así que el guardián
  `IF($C5="","",SUMIF(...,$C5,...))` NO protege: en las filas libres el criterio
  llega vacío, `sumif` revienta con «Couldn't parse criteria: None» y el libro
  ENTERO se queda sin cache (medido en este mismo producto). Los agregados por
  familia van con `SUMPRODUCT(--(rango=celda), …)`, que sí tolera el vacío — que
  es exactamente lo que hace el `menu-engineering-matrix.xlsx` de la guía
  gastronómica. `SUMIF`/`COUNTIF` sólo se usan donde el criterio es una etiqueta
  FIJA (resúmenes por familia y por clasificación).
* Funciones prohibidas (INDIRECT, COUNTA, PMT, OFFSET, LET, LAMBDA, XLOOKUP,
  matrices dinámicas): cero. Tampoco `COUNT`, que no está en la lista blanca —
  su papel lo hace `COUNTIF(rango,"<>")` o `SUMPRODUCT`.
* Cero constantes dentro de fórmulas: tipo de IVA, factor del umbral de
  popularidad y porcentajes de personal y otros variables del Goal Value viven
  en celdas verdes con su nota y se referencian con `$` absolutos.
* «Sin dato» = `""`, nunca `0`; `IFERROR(...,"")` en todo cociente.

Salida fija (sin argumentos): `<carpeta>/build/matriz-multimetodo-carta.xlsx`
"""
import json
import os
import sys

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

motor.CTX['producto'] = 'guia-food-cost-ingenieria-menu'

# --------------------------------------------------------------------------
# Constantes de familia
# --------------------------------------------------------------------------
NOMBRE = 'matriz-multimetodo-carta'
TITULO_LIBRO = 'Matriz Multi-Método de la Carta'
SUBTITULO = 'AI Chef Pro · aichef.pro — Guía Food Cost + Ingeniería de Menú'
SUBJECT = 'Guía Food Cost + Ingeniería de Menú · Versión 1.0 · septiembre 2026'
VERSION = ('Versión 1.0 · septiembre 2026 · '
           'aichef.pro/guia-food-cost-ingenieria-menu · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 · '
       'johnguerrero.es')
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde, '
               'desprotege la hoja (sin contraseña).')
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

N = motor.NARROW                       # espacio fino antes de la unidad
GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL = 'FFF8E1', '1565C0'
EUR, PCT, PCT0, ENT = motor.FMT_EUR, motor.FMT_PCT, '0%', motor.FMT_ENT
IDX = '#,##0.00'   # el Goal Value es un ÍNDICE sin unidades, no euros

FAMILIAS = ['Entrantes', 'Principales', 'Postres', 'Menú']

FIL0, FIL1 = 5, 29                     # 20 platos de ejemplo + 5 filas libres

# Nombre de hoja SIEMPRE entrecomillado: «Kasavana-Smith» y «Goal Value» no son
# identificadores válidos y sin comillas Excel los leería como una resta.
DAT = "'Datos'!"
KS = "'Kasavana-Smith'!"
MIL = "'Miller'!"
PAV = "'Pavesic'!"
GVA = "'Goal Value'!"

P_UMBRAL = '$D$33'        # Datos: factor del umbral de popularidad
P_IVA = '$D$34'           # Datos: tipo de IVA de restauración en sala
P_LABOR = '$D$33'         # Goal Value: coste de personal sobre ventas
P_VAR = '$D$34'           # Goal Value: otros costes variables sobre ventas


# --------------------------------------------------------------------------
# Utilidades de hoja
# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, landscape=True):
    ws.page_setup.paperSize = 9                     # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8


def encabezados(ws, fila, cols, alto=46):
    """Cabecera: fondo 2D2D2D, letra blanca en negrita, texto ajustado.

    `ancho` a None deja el ancho que ya tenga la columna (los bloques de
    resumen comparten columnas con la tabla principal y no deben pisarlo).
    """
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def bloque(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)


def total(ws, coord, contenido, fmt=None, formula=False):
    cel = (motor.f(ws, coord, contenido, fmt=fmt, bold=True) if formula
           else motor.val(ws, coord, contenido, fmt=fmt, bold=True))
    cel.fill = PatternFill('solid', fgColor=CREMA)
    return cel


def cf_reglas(ws, rango, reglas):
    """Varias reglas de expresión sobre el MISMO rango.

    `motor.regla_expresion` purga las reglas cuyo sqref coincide antes de
    añadir la suya: encadenar dos llamadas borraría la primera.
    """
    motor._limpiar_cf(ws, rango)
    for formula, bg, fg in reglas:
        ws.conditional_formatting.add(
            rango, FormulaRule(formula=[formula], stopIfTrue=True,
                               font=Font(color=fg, bold=True),
                               fill=PatternFill(start_color=bg, end_color=bg,
                                                fill_type='solid')))


def parametro(ws, fila, col_et, col_val, clave, col_nota=None):
    """Parámetro normativo de la familia: etiqueta + celda VERDE + nota."""
    p = motor.PARAMETROS[clave]
    motor.val(ws, col_et + str(fila), p['etiqueta'])
    motor.val(ws, col_val + str(fila), p['valor'], fmt=p['formato'],
              verde_=True)
    if col_nota:
        motor.val(ws, col_nota + str(fila), p['nota'])
    return '$' + col_val + '$' + str(fila)


def verde_propio(ws, fila, col_et, col_val, etiqueta, valor, fmt, nota,
                 col_nota=None):
    motor.val(ws, col_et + str(fila), etiqueta)
    motor.val(ws, col_val + str(fila), valor, fmt=fmt, verde_=True)
    if col_nota:
        motor.val(ws, col_nota + str(fila), nota)
    return '$' + col_val + '$' + str(fila)


def espejo(ws, col, fila, origen_col, fmt=None, bold=None):
    """Columna que replica un dato de «Datos» para leer la hoja por su cuenta."""
    return motor.f(ws, col + str(fila),
                   f'=IFERROR(IF({DAT}${origen_col}{fila}="","",'
                   f'{DAT}${origen_col}{fila}),"")',
                   fmt=fmt, bold=bold)


def cadena_if(col, pares, defecto):
    """`IF` encadenados sobre una columna de clasificación."""
    out = ''
    for valor, texto in pares:
        out += f'IF({col}="{valor}","{texto}",'
    return out + f'"{defecto}"' + ')' * len(pares)


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. Vuelca tu carta en la hoja «Datos»: plato, familia, unidades vendidas '
    'del último mes, coste por ración y precio de venta SIN IVA.',
    '2. El coste por ración sale de tu ficha de escandallo (libro '
    '«ficha-escandallo-base» de este mismo pack). El PVP con IVA de sala lo '
    'calcula el libro.',
    '3. Las hojas Kasavana-Smith, Miller, Pavesic y Goal Value leen esos datos '
    'y clasifican cada plato con su método. En ellas no se teclea nada.',
    '4. Cada método se aplica DENTRO de su familia: un postre se compara con '
    'los postres, no con los chuletones.',
    '5. Abre «Comparativa»: ahí tienes las cuatro lecturas del mismo plato, '
    'cuántas se salen de la mejor categoría y un diagnóstico de qué hacer.',
    '6. Lleva las decisiones al libro «plan-accion-90-dias» con responsable y '
    'fecha. Un análisis sin fecha no cambia nada.',
    '7. La hoja «Menú Precio Fijo» va aparte: en un menú cerrado el margen no '
    'lo decide el precio, lo decide el mix de elecciones de tus comensales.',
    '8. Ajusta el factor del umbral de popularidad y los porcentajes de '
    'personal y de otros costes variables si tu casa no se parece al ejemplo.',
]

NOTAS = [
    'Los datos de ejemplo son de un restaurante MODELADO, no de un cliente '
    'real. Bórralos y pon los tuyos: hay 20 platos rellenos y 5 filas libres.',
    'Kasavana & Smith cruza popularidad con margen de contribución en euros. '
    'Miller cruza popularidad con food cost' + N + '%. Pavesic cruza food '
    'cost' + N + '% con el margen ponderado por unidades. Goal Value (Hayes & '
    'Huffman) añade el peso del personal y de los otros costes variables '
    'SIN dejar de contar el food cost' + N + '%: por eso su factor final es '
    '(1 − (personal' + N + '% + otros variables' + N + '% + food cost' + N +
    '%)) y no sólo personal + otros. Ninguno mide las tres variables a la '
    'vez: por eso discrepan.',
    'El umbral de popularidad clásico es el 70' + N + '% del mix medio: '
    'umbral = factor ÷ número de platos de la familia con ventas. Súbelo si la '
    'familia es muy corta.',
    'Que los cuatro métodos coincidan NO significa «más seguridad»: significa '
    'que el plato es tan bueno o tan malo que se ve desde cualquier ángulo. Lo '
    'interesante está en las filas donde discrepan.',
    'El margen de contribución es el euro que queda para pagar personal, '
    'alquiler y luz. El food cost' + N + '% mide eficiencia. Un plato puede ser '
    'eficiente y aportar poco dinero, y al revés.',
]

NOTA_IVA_LIBRO = (
    'El precio de venta de la hoja «Datos» va SIN IVA. La columna «PVP con IVA '
    'en sala» aplica el 10' + N + '% del art. 91.Uno.2.2.º de la Ley del IVA, '
    'que en sala cubre también la bebida alcohólica. Para llevar y en delivery '
    'el tipo cambia: eso se calcula en el libro '
    '«simulador-repricing-multicanal» de este pack.')


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila), NOTA_IVA_LIBRO, wrap=True)
    fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, landscape=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Datos»
# --------------------------------------------------------------------------
COLS_DATOS = [
    ('A', '#', 5),
    ('B', 'Plato', 44),
    ('C', 'Familia', 14),
    ('D', 'Uds vendidas (mes)', 12),
    ('E', 'Coste por ración (€)', 13),
    ('F', 'PVP sin IVA (€)', 12),
    ('G', 'Margen de contribución (€)', 14),
    ('H', 'Food cost (%)', 11),
    ('I', 'PVP con IVA en sala (€)', 13),
    ('J', 'Uds de su familia', 12),
    ('K', 'Platos de su familia con ventas', 13),
    ('L', 'Mix dentro de su familia (%)', 13),
    ('M', 'Umbral de popularidad de su familia (%)', 14),
    ('N', 'MC medio ponderado de su familia (€)', 15),
    ('O', 'Food cost medio ponderado de su familia (%)', 15),
    ('P', 'MC ponderado del plato: MC × uds (€)', 15),
    ('Q', 'Media del MC ponderado de su familia (€)', 16),
]


def hoja_datos(wb):
    ws = wb.create_sheet('Datos')
    cabecera(ws, 'Datos de la carta')
    apunte(ws, 'E3', 'El coste por ración sale de tu ficha de escandallo; el '
                     'PVP va SIN IVA.')
    encabezados(ws, 4, COLS_DATOS)
    ws.freeze_panes = 'B5'

    v_lista, v_ent, v_eur = [], [], []
    for i in range(FIL1 - FIL0 + 1):
        r = FIL0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        if i < len(DE.PLATOS):
            _id, nombre, familia, coste, pvp, uds = DE.PLATOS[i]
            motor.val(ws, f'B{r}', nombre)
            motor.val(ws, f'C{r}', familia)
            motor.val(ws, f'D{r}', uds, fmt=ENT)
            motor.val(ws, f'E{r}', coste, fmt=EUR)
            motor.val(ws, f'F{r}', pvp, fmt=EUR)
        else:
            ws[f'D{r}'].number_format = ENT
            ws[f'E{r}'].number_format = EUR
            ws[f'F{r}'].number_format = EUR
        motor.verde(ws, f'B{r}:F{r}')
        v_lista.append(f'C{r}')
        v_ent.append(f'D{r}')
        v_eur += [f'E{r}', f'F{r}']

        fam = f'--($C${FIL0}:$C${FIL1}=$C{r})'
        uds_r = f'$D${FIL0}:$D${FIL1}'
        mc_r = f'$G${FIL0}:$G${FIL1}'
        cos_r = f'$E${FIL0}:$E${FIL1}'
        pvp_r = f'$F${FIL0}:$F${FIL1}'

        motor.f(ws, f'G{r}',
                f'=IFERROR(IF(OR($E{r}="",$F{r}=""),"",$F{r}-$E{r}),"")',
                fmt=EUR)
        ws[f'G{r}'].font = Font(bold=True, color=AZUL)
        motor.f(ws, f'H{r}',
                f'=IFERROR(IF(OR($E{r}="",$F{r}="",$F{r}=0),"",$E{r}/$F{r}),"")',
                fmt=PCT)
        ws[f'H{r}'].font = Font(bold=True, color=AZUL)
        motor.f(ws, f'I{r}',
                f'=IFERROR(IF(OR($F{r}="",{P_IVA}=""),"",$F{r}*(1+{P_IVA})),"")',
                fmt=EUR)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF($C{r}="","",SUMPRODUCT({fam},{uds_r})),"")',
                fmt=ENT)
        motor.f(ws, f'K{r}',
                f'=IFERROR(IF($C{r}="","",SUMPRODUCT({fam},--({uds_r}>0))),"")',
                fmt=ENT)
        motor.f(ws, f'L{r}',
                f'=IFERROR(IF(OR($D{r}="",$J{r}="",$J{r}=0),"",$D{r}/$J{r}),"")',
                fmt=PCT)
        motor.f(ws, f'M{r}',
                f'=IFERROR(IF(OR($K{r}="",$K{r}=0,{P_UMBRAL}=""),"",'
                f'{P_UMBRAL}/$K{r}),"")', fmt=PCT)
        motor.f(ws, f'N{r}',
                f'=IFERROR(IF(OR($J{r}="",$J{r}=0),"",'
                f'SUMPRODUCT({fam},{uds_r},{mc_r})/$J{r}),"")', fmt=EUR)
        motor.f(ws, f'O{r}',
                f'=IFERROR(IF(SUMPRODUCT({fam},{uds_r},{pvp_r})=0,"",'
                f'SUMPRODUCT({fam},{uds_r},{cos_r})/'
                f'SUMPRODUCT({fam},{uds_r},{pvp_r})),"")', fmt=PCT)
        motor.f(ws, f'P{r}',
                f'=IFERROR(IF(OR($G{r}="",$D{r}=""),"",$G{r}*$D{r}),"")',
                fmt=EUR)
        motor.f(ws, f'Q{r}',
                f'=IFERROR(IF(OR($K{r}="",$K{r}=0),"",'
                f'SUMPRODUCT({fam},{uds_r},{mc_r})/$K{r}),"")', fmt=EUR)

    motor.dv_lista(ws, v_lista, FAMILIAS, titulo='Familia no válida')
    motor.dv_numerica(ws, v_ent, minimo=0, titulo='Unidades no válidas',
                      mensaje='Escribe las unidades vendidas del último mes '
                              '(0 o más).')
    motor.dv_numerica(ws, v_eur, minimo=0)

    # --- Resumen y parámetros -------------------------------------------
    bloque(ws, 'A31', 'RESUMEN Y PARÁMETROS — lo calcula el libro')
    total(ws, 'B32', 'TOTAL DE LA CARTA')
    total(ws, 'D32', f'=IFERROR(SUM($D${FIL0}:$D${FIL1}),"")', fmt=ENT,
          formula=True)
    total(ws, 'G32',
          f'=IFERROR(IF($D$32=0,"",SUMPRODUCT($D${FIL0}:$D${FIL1},'
          f'$G${FIL0}:$G${FIL1})/$D$32),"")', fmt=EUR, formula=True)
    total(ws, 'H32',
          f'=IFERROR(IF(SUMPRODUCT($D${FIL0}:$D${FIL1},$F${FIL0}:$F${FIL1})=0,'
          f'"",SUMPRODUCT($D${FIL0}:$D${FIL1},$E${FIL0}:$E${FIL1})/'
          f'SUMPRODUCT($D${FIL0}:$D${FIL1},$F${FIL0}:$F${FIL1})),"")',
          fmt=PCT, formula=True)
    total(ws, 'I32',
          f'=IFERROR(SUMPRODUCT($D${FIL0}:$D${FIL1},$F${FIL0}:$F${FIL1}),"")',
          fmt=EUR, formula=True)
    apunte(ws, 'J32', 'Uds · MC medio ponderado · food cost medio ponderado · '
                      'ventas netas del mes')

    verde_propio(ws, 33, 'B', 'D', 'Factor del umbral de popularidad (%)',
                 0.7, PCT0,
                 'El clásico es el 70' + N + '% del mix medio: el umbral de '
                 'cada familia = este factor ÷ número de platos de la familia '
                 'con ventas. Súbelo si la familia es muy corta.',
                 col_nota='F')
    motor.dv_porcentaje(ws, ['D33'], titulo='Factor del umbral',
                        prompt='Se escribe en tanto por uno: 0,70 = 70' + N +
                               '%.')
    parametro(ws, 34, 'B', 'D', 'iva_restauracion', col_nota='F')
    motor.dv_porcentaje(ws, ['D34'], titulo='Tipo de IVA',
                        prompt='Se escribe en tanto por uno: 0,10 = 10' + N +
                               '%.')
    motor.val(ws, 'B35', 'Platos con ventas en la carta')
    motor.f(ws, 'D35', f'=COUNTIF($D${FIL0}:$D${FIL1},">0")', fmt=ENT)
    motor.val(ws, 'B36', 'Platos dados de alta en la carta')
    motor.f(ws, 'D36', f'=COUNTIF($B${FIL0}:$B${FIL1},"<>")', fmt=ENT)

    # --- Resumen por familia ---------------------------------------------
    bloque(ws, 'A38', 'RESUMEN POR FAMILIA — lo calcula el libro')
    encabezados(ws, 39, [
        ('B', 'Familia', None), ('C', 'Platos en carta', None),
        ('D', 'Platos con ventas', None), ('E', 'Uds vendidas', None),
        ('F', 'Mix sobre la carta (%)', None),
        ('G', 'MC medio ponderado (€)', None),
        ('H', 'Food cost medio ponderado (%)', None),
        ('I', 'Ventas netas del mes (€)', None),
    ], alto=34)
    for k, fam in enumerate(FAMILIAS):
        r = 40 + k
        motor.val(ws, f'B{r}', fam, bold=True)
        sel = f'--($C${FIL0}:$C${FIL1}=$B{r})'
        motor.f(ws, f'C{r}', f'=COUNTIF($C${FIL0}:$C${FIL1},$B{r})', fmt=ENT)
        motor.f(ws, f'D{r}',
                f'=IFERROR(IF($C{r}=0,"",SUMPRODUCT({sel},'
                f'--($D${FIL0}:$D${FIL1}>0))),"")', fmt=ENT)
        motor.f(ws, f'E{r}',
                f'=IFERROR(IF($C{r}=0,"",SUMIF($C${FIL0}:$C${FIL1},$B{r},'
                f'$D${FIL0}:$D${FIL1})),"")', fmt=ENT)
        motor.f(ws, f'F{r}',
                f'=IFERROR(IF(OR($D$32=0,$C{r}=0),"",$E{r}/$D$32),"")',
                fmt=PCT)
        motor.f(ws, f'G{r}',
                f'=IFERROR(IF($E{r}=0,"",SUMPRODUCT({sel},$D${FIL0}:$D${FIL1},'
                f'$G${FIL0}:$G${FIL1})/$E{r}),"")', fmt=EUR)
        motor.f(ws, f'H{r}',
                f'=IFERROR(IF(SUMPRODUCT({sel},$D${FIL0}:$D${FIL1},'
                f'$F${FIL0}:$F${FIL1})=0,"",SUMPRODUCT({sel},'
                f'$D${FIL0}:$D${FIL1},$E${FIL0}:$E${FIL1})/SUMPRODUCT({sel},'
                f'$D${FIL0}:$D${FIL1},$F${FIL0}:$F${FIL1})),"")', fmt=PCT)
        motor.f(ws, f'I{r}',
                f'=IFERROR(SUMPRODUCT({sel},$D${FIL0}:$D${FIL1},'
                f'$F${FIL0}:$F${FIL1}),"")', fmt=EUR)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Bloque de resumen por clasificación (K&S, Miller, Pavesic)
# --------------------------------------------------------------------------
def resumen_clases(ws, fila_titulo, clases, col_clase, col_uds, extra):
    """extra = (cabecera, formato, plantilla con {r} y el rango ya escrito)."""
    cab_extra, fmt_extra, plantilla = extra
    bloque(ws, f'A{fila_titulo}', 'RESUMEN — lo calcula el libro')
    fc = fila_titulo + 1
    encabezados(ws, fc, [
        ('B', 'Clasificación', None), ('C', 'Platos', None),
        ('D', 'Uds vendidas', None), ('E', '% de las uds de la carta', None),
        ('F', cab_extra, None),
    ], alto=34)
    for k, clase in enumerate(clases):
        r = fc + 1 + k
        motor.val(ws, f'B{r}', clase, bold=True)
        motor.f(ws, f'C{r}',
                f'=COUNTIF(${col_clase}${FIL0}:${col_clase}${FIL1},$B{r})',
                fmt=ENT)
        motor.f(ws, f'D{r}',
                f'=IFERROR(SUMIF(${col_clase}${FIL0}:${col_clase}${FIL1},$B{r},'
                f'${col_uds}${FIL0}:${col_uds}${FIL1}),"")', fmt=ENT)
        motor.f(ws, f'E{r}',
                f'=IFERROR(IF(SUM(${col_uds}${FIL0}:${col_uds}${FIL1})=0,"",'
                f'$D{r}/SUM(${col_uds}${FIL0}:${col_uds}${FIL1})),"")',
                fmt=PCT)
        motor.f(ws, f'F{r}', plantilla.format(r=r), fmt=fmt_extra)
    return fc + len(clases)


# --------------------------------------------------------------------------
# Kasavana & Smith
# --------------------------------------------------------------------------
KS_SIGNIFICA = [
    ('Star', 'Se vende y deja margen: es el plato que sostiene a su familia.'),
    ('Plowhorse', 'Se vende mucho, pero deja poco margen en euros por unidad.'),
    ('Puzzle', 'Deja buen margen en euros, pero se pide poco.'),
]
KS_ACCION = [
    ('Star', 'Mantener receta y proveedor, y darle sitio destacado en la carta.'),
    ('Plowhorse', 'Subir precio con cuidado o bajar coste sin tocar la ración.'),
    ('Puzzle', 'Reubicar en la carta, renombrar y que la sala lo sugiera.'),
]


def hoja_kasavana(wb):
    ws = wb.create_sheet('Kasavana-Smith')
    cabecera(ws, 'Kasavana & Smith — popularidad × margen de contribución')
    apunte(ws, 'E3', 'Todo se calcula desde la hoja «Datos». Aquí no se teclea '
                     'nada.')
    encabezados(ws, 4, [
        ('A', '#', 5), ('B', 'Plato', 44), ('C', 'Familia', 14),
        ('D', 'Uds vendidas', 12), ('E', 'Mix dentro de su familia (%)', 13),
        ('F', 'Umbral de popularidad de su familia (%)', 14),
        ('G', 'Popularidad', 12), ('H', 'MC del plato (€)', 12),
        ('I', 'MC medio ponderado de su familia (€)', 15),
        ('J', 'Nivel de margen', 12), ('K', 'Clasificación', 14),
        ('L', 'Qué significa', 46), ('M', 'Acción recomendada', 42),
    ])
    ws.freeze_panes = 'B5'
    for i in range(FIL1 - FIL0 + 1):
        r = FIL0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        espejo(ws, 'B', r, 'B')
        espejo(ws, 'C', r, 'C')
        espejo(ws, 'D', r, 'D', fmt=ENT)
        espejo(ws, 'E', r, 'L', fmt=PCT)
        espejo(ws, 'F', r, 'M', fmt=PCT)
        motor.f(ws, f'G{r}',
                f'=IFERROR(IF(OR($E{r}="",$F{r}=""),"",'
                f'IF($E{r}>=$F{r},"Alta","Baja")),"")')
        espejo(ws, 'H', r, 'G', fmt=EUR)
        ws[f'H{r}'].font = Font(bold=True, color=AZUL)
        espejo(ws, 'I', r, 'N', fmt=EUR)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF(OR($H{r}="",$I{r}=""),"",'
                f'IF($H{r}>=$I{r},"Alto","Bajo")),"")')
        motor.f(ws, f'K{r}',
                f'=IFERROR(IF(OR($G{r}="",$J{r}=""),"",'
                f'IF(AND($G{r}="Alta",$J{r}="Alto"),"Star",'
                f'IF(AND($G{r}="Alta",$J{r}="Bajo"),"Plowhorse",'
                f'IF(AND($G{r}="Baja",$J{r}="Alto"),"Puzzle","Dog")))),"")',
                bold=True)
        motor.f(ws, f'L{r}',
                f'=IFERROR(IF($K{r}="","",' + cadena_if(
                    f'$K{r}', KS_SIGNIFICA,
                    'Ni se vende ni deja margen.') + '),"")')
        motor.f(ws, f'M{r}',
                f'=IFERROR(IF($K{r}="","",' + cadena_if(
                    f'$K{r}', KS_ACCION,
                    'Retirar o reformular: ocupa sitio en la carta y en la '
                    'cocina.') + '),"")')
    motor.semaforo_texto(ws, f'K{FIL0}:K{FIL1}', (
        ('Star', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Plowhorse', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Puzzle', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Dog', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    resumen_clases(
        ws, 31, ['Star', 'Plowhorse', 'Puzzle', 'Dog'], 'K', 'D',
        ('MC total aportado (€)', EUR,
         '=IFERROR(SUMPRODUCT(--($K$' + str(FIL0) + ':$K$' + str(FIL1) +
         '=$B{r}),$D$' + str(FIL0) + ':$D$' + str(FIL1) + ',$H$' + str(FIL0) +
         ':$H$' + str(FIL1) + '),"")'))
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Miller
# --------------------------------------------------------------------------
MIL_SIGNIFICA = [
    ('Winner', 'Popular y con el food cost por debajo de la media de su '
               'familia.'),
    ('Marginal', 'Cumple en una sola de las dos: o vende, o tiene el food cost '
                 'sano.'),
]
MIL_ACCION = [
    ('Winner', 'Mantener y proteger el escandallo: es el que aguanta una '
               'subida del proveedor.'),
    ('Marginal', 'Mirar las otras lecturas antes de tocarlo: el problema es de '
                 'precio o es de carta.'),
]


def hoja_miller(wb):
    ws = wb.create_sheet('Miller')
    cabecera(ws, 'Miller — popularidad × food cost %')
    apunte(ws, 'E3', 'Miller premia el food cost bajo; Kasavana & Smith premia '
                     'el margen en euros. No siempre coinciden.')
    encabezados(ws, 4, [
        ('A', '#', 5), ('B', 'Plato', 44), ('C', 'Familia', 14),
        ('D', 'Uds vendidas', 12), ('E', 'Mix dentro de su familia (%)', 13),
        ('F', 'Umbral de popularidad de su familia (%)', 14),
        ('G', 'Popularidad', 12), ('H', 'Food cost del plato (%)', 12),
        ('I', 'Food cost medio ponderado de su familia (%)', 15),
        ('J', 'Nivel de food cost', 12), ('K', 'Clasificación', 14),
        ('L', 'Qué significa', 46), ('M', 'Acción recomendada', 42),
    ])
    ws.freeze_panes = 'B5'
    for i in range(FIL1 - FIL0 + 1):
        r = FIL0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        espejo(ws, 'B', r, 'B')
        espejo(ws, 'C', r, 'C')
        espejo(ws, 'D', r, 'D', fmt=ENT)
        espejo(ws, 'E', r, 'L', fmt=PCT)
        espejo(ws, 'F', r, 'M', fmt=PCT)
        motor.f(ws, f'G{r}',
                f'=IFERROR(IF(OR($E{r}="",$F{r}=""),"",'
                f'IF($E{r}>=$F{r},"Alta","Baja")),"")')
        espejo(ws, 'H', r, 'H', fmt=PCT)
        ws[f'H{r}'].font = Font(bold=True, color=AZUL)
        espejo(ws, 'I', r, 'O', fmt=PCT)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF(OR($H{r}="",$I{r}=""),"",'
                f'IF($H{r}<=$I{r},"Bajo","Alto")),"")')
        motor.f(ws, f'K{r}',
                f'=IFERROR(IF(OR($G{r}="",$J{r}=""),"",'
                f'IF(AND($G{r}="Alta",$J{r}="Bajo"),"Winner",'
                f'IF(AND($G{r}="Baja",$J{r}="Alto"),"Loser","Marginal"))),"")',
                bold=True)
        motor.f(ws, f'L{r}',
                f'=IFERROR(IF($K{r}="","",' + cadena_if(
                    f'$K{r}', MIL_SIGNIFICA,
                    'Poco pedido y con el food cost por encima de su '
                    'familia.') + '),"")')
        motor.f(ws, f'M{r}',
                f'=IFERROR(IF($K{r}="","",' + cadena_if(
                    f'$K{r}', MIL_ACCION,
                    'Reformular la receta o retirarlo de la carta.') + '),"")')
    motor.semaforo_texto(ws, f'K{FIL0}:K{FIL1}', (
        ('Winner', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Marginal', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Loser', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    resumen_clases(
        ws, 31, ['Winner', 'Marginal', 'Loser'], 'K', 'D',
        ('Food cost medio ponderado del grupo (%)', PCT,
         '=IFERROR(IF($D{r}=0,"",SUMPRODUCT(--($K$' + str(FIL0) + ':$K$' +
         str(FIL1) + '=$B{r}),$D$' + str(FIL0) + ':$D$' + str(FIL1) + ',$H$' +
         str(FIL0) + ':$H$' + str(FIL1) + ')/$D{r}),"")'))
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Pavesic
# --------------------------------------------------------------------------
PAV_SIGNIFICA = [
    ('Prime', 'Food cost bajo y aporta mucho margen en euros al total.'),
    ('Standard', 'Aporta mucho margen en euros pese a tener el food cost alto.'),
    ('Sleeper', 'Food cost sano, pero aporta poco margen en euros al total.'),
]
PAV_ACCION = [
    ('Prime', 'Mantener y promocionar: aquí está el dinero de la familia.'),
    ('Standard', 'Vigilar el escandallo: gana por volumen, no por eficiencia.'),
    ('Sleeper', 'Darle visibilidad: el margen está, falta que se pida.'),
]


def hoja_pavesic(wb):
    ws = wb.create_sheet('Pavesic')
    cabecera(ws, 'Pavesic — food cost % × margen ponderado por unidades')
    apunte(ws, 'E3', 'El margen ponderado (MC × uds) mide cuánto dinero aporta '
                     'el plato al total, no cuánto deja por unidad.')
    encabezados(ws, 4, [
        ('A', '#', 5), ('B', 'Plato', 44), ('C', 'Familia', 14),
        ('D', 'Uds vendidas', 12), ('E', 'Food cost del plato (%)', 12),
        ('F', 'Food cost medio ponderado de su familia (%)', 15),
        ('G', 'Nivel de food cost', 12),
        ('H', 'MC ponderado: MC × uds (€)', 14),
        ('I', 'Media del MC ponderado de su familia (€)', 15),
        ('J', 'Nivel de MC ponderado', 13), ('K', 'Clasificación', 14),
        ('L', 'Qué significa', 46), ('M', 'Acción recomendada', 42),
    ])
    ws.freeze_panes = 'B5'
    for i in range(FIL1 - FIL0 + 1):
        r = FIL0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        espejo(ws, 'B', r, 'B')
        espejo(ws, 'C', r, 'C')
        espejo(ws, 'D', r, 'D', fmt=ENT)
        espejo(ws, 'E', r, 'H', fmt=PCT)
        espejo(ws, 'F', r, 'O', fmt=PCT)
        motor.f(ws, f'G{r}',
                f'=IFERROR(IF(OR($E{r}="",$F{r}=""),"",'
                f'IF($E{r}<=$F{r},"Bajo","Alto")),"")')
        espejo(ws, 'H', r, 'P', fmt=EUR)
        ws[f'H{r}'].font = Font(bold=True, color=AZUL)
        espejo(ws, 'I', r, 'Q', fmt=EUR)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF(OR($H{r}="",$I{r}=""),"",'
                f'IF($H{r}>=$I{r},"Alto","Bajo")),"")')
        motor.f(ws, f'K{r}',
                f'=IFERROR(IF(OR($G{r}="",$J{r}=""),"",'
                f'IF(AND($G{r}="Bajo",$J{r}="Alto"),"Prime",'
                f'IF(AND($G{r}="Alto",$J{r}="Alto"),"Standard",'
                f'IF(AND($G{r}="Bajo",$J{r}="Bajo"),"Sleeper","Problem")))),"")',
                bold=True)
        motor.f(ws, f'L{r}',
                f'=IFERROR(IF($K{r}="","",' + cadena_if(
                    f'$K{r}', PAV_SIGNIFICA,
                    'Food cost alto y aporta poco margen en euros al '
                    'total.') + '),"")')
        motor.f(ws, f'M{r}',
                f'=IFERROR(IF($K{r}="","",' + cadena_if(
                    f'$K{r}', PAV_ACCION,
                    'Reformular la receta o retirarlo de la carta.') + '),"")')
    motor.semaforo_texto(ws, f'K{FIL0}:K{FIL1}', (
        ('Prime', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Standard', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Sleeper', motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('Problem', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    resumen_clases(
        ws, 31, ['Prime', 'Standard', 'Sleeper', 'Problem'], 'K', 'D',
        ('MC ponderado total del grupo (€)', EUR,
         '=IFERROR(SUMIF($K$' + str(FIL0) + ':$K$' + str(FIL1) + ',$B{r},$H$' +
         str(FIL0) + ':$H$' + str(FIL1) + '),"")'))
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Goal Value (Hayes & Huffman)
# --------------------------------------------------------------------------
def hoja_goal_value(wb):
    ws = wb.create_sheet('Goal Value')
    cabecera(ws, 'Goal Value (Hayes & Huffman) — el método que mete el personal')
    apunte(ws, 'E3', 'Goal Value = (1 − food cost' + N + '%) × uds × PVP × '
                     '(1 − (personal' + N + '% + otros variables' + N +
                     '% + food cost' + N + '%)) — el food cost pesa dos veces: '
                     'es lo que separa a Goal Value de Pavesic y penaliza al '
                     'plato caro de producir.')
    encabezados(ws, 4, [
        ('A', '#', 5), ('B', 'Plato', 44), ('C', 'Familia', 14),
        ('D', 'Uds vendidas', 12), ('E', 'Food cost del plato (%)', 12),
        ('F', 'PVP sin IVA (€)', 12), ('G', 'Goal Value del plato', 14),
        ('H', 'Uds medias de su familia', 13),
        ('I', 'PVP medio ponderado de su familia (€)', 15),
        ('J', 'Food cost medio ponderado de su familia (%)', 15),
        ('K', 'Goal Value objetivo de su familia', 15),
        ('L', 'Lectura', 22), ('M', 'Distancia al objetivo (índice)', 14),
        ('N', 'Qué revisar', 62),
    ])
    ws.freeze_panes = 'B5'
    for i in range(FIL1 - FIL0 + 1):
        r = FIL0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        espejo(ws, 'B', r, 'B')
        espejo(ws, 'C', r, 'C')
        espejo(ws, 'D', r, 'D', fmt=ENT)
        espejo(ws, 'E', r, 'H', fmt=PCT)
        espejo(ws, 'F', r, 'F', fmt=EUR)
        motor.f(ws, f'G{r}',
                f'=IFERROR(IF(OR($D{r}="",$E{r}="",$F{r}="",'
                f'{P_LABOR}+{P_VAR}+$E{r}>=1),"",'
                f'(1-$E{r})*$D{r}*$F{r}*(1-({P_LABOR}+{P_VAR}+$E{r}))),"")',
                fmt=IDX)
        ws[f'G{r}'].font = Font(bold=True, color=AZUL)
        motor.f(ws, f'H{r}',
                f'=IFERROR(IF(OR({DAT}$K{r}="",{DAT}$K{r}=0),"",'
                f'{DAT}$J{r}/{DAT}$K{r}),"")', fmt=ENT)
        motor.f(ws, f'I{r}',
                f'=IFERROR(IF(OR({DAT}$J{r}="",{DAT}$J{r}=0),"",'
                f'SUMPRODUCT(--({DAT}$C${FIL0}:$C${FIL1}={DAT}$C{r}),'
                f'{DAT}$D${FIL0}:$D${FIL1},{DAT}$F${FIL0}:$F${FIL1})/'
                f'{DAT}$J{r}),"")', fmt=EUR)
        espejo(ws, 'J', r, 'O', fmt=PCT)
        motor.f(ws, f'K{r}',
                f'=IFERROR(IF(OR($H{r}="",$I{r}="",$J{r}="",'
                f'{P_LABOR}+{P_VAR}+$J{r}>=1),"",'
                f'(1-$J{r})*$H{r}*$I{r}*(1-({P_LABOR}+{P_VAR}+$J{r}))),"")',
                fmt=IDX)
        motor.f(ws, f'L{r}',
                f'=IFERROR(IF(OR($G{r}="",$K{r}=""),"",'
                f'IF($G{r}>=$K{r},"Por encima del objetivo",'
                f'"Por debajo del objetivo")),"")', bold=True)
        motor.f(ws, f'M{r}',
                f'=IFERROR(IF(OR($G{r}="",$K{r}=""),"",$G{r}-$K{r}),"")',
                fmt=IDX)
        motor.f(ws, f'N{r}',
                f'=IFERROR(IF($L{r}="","",'
                f'IF($L{r}="Por encima del objetivo",'
                f'"Cumple el objetivo de su familia: mantenlo y protégelo.",'
                f'IF($E{r}>$J{r},'
                f'"Su food cost está por encima de la media de la familia: '
                f'baja coste o sube precio.",'
                f'IF($D{r}<$H{r},'
                f'"Vende menos que la media de su familia: es problema de '
                f'carta y de sala, no de coste.",'
                f'"Su precio está por debajo de la media de su familia: '
                f'revísalo antes de tocar la receta.")))),"")')
    motor.semaforo_texto(ws, f'L{FIL0}:L{FIL1}', (
        ('Por encima del objetivo', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Por debajo del objetivo', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))

    bloque(ws, 'A32', 'PARÁMETROS Y RESUMEN — lo calcula el libro')
    verde_propio(ws, 33, 'B', 'D', 'Coste de personal sobre ventas (%)',
                 0.32, PCT,
                 'Personal con Seguridad Social incluida, sobre ventas netas. '
                 'El 32' + N + '% es el punto medio de la horquilla 30-35' + N +
                 '% del servicio en mesa en España; en barra o autoservicio '
                 'baja. Cámbialo por el tuyo.', col_nota='F')
    verde_propio(ws, 34, 'B', 'D', 'Otros costes variables sobre ventas (%)',
                 0.10, PCT,
                 'Lo que se mueve con la venta y no es producto ni personal: '
                 'energía de cocina, consumibles y comisiones de tarjeta y de '
                 'plataformas. Cámbialo por el tuyo.', col_nota='F')
    motor.dv_porcentaje(ws, ['D33', 'D34'], titulo='Porcentaje sobre ventas',
                        prompt='Se escribe en tanto por uno: 0,32 = 32' + N +
                               '%.')
    motor.val(ws, 'B35', 'Goal Value medio de la carta (índice)')
    motor.f(ws, 'D35', f'=IFERROR(AVERAGE($G${FIL0}:$G${FIL1}),"")', fmt=IDX)
    motor.val(ws, 'B36', 'Platos por encima del objetivo de su familia')
    motor.f(ws, 'D36',
            f'=COUNTIF($L${FIL0}:$L${FIL1},"Por encima del objetivo")', fmt=ENT)
    motor.val(ws, 'B37', 'Platos por debajo del objetivo de su familia')
    motor.f(ws, 'D37',
            f'=COUNTIF($L${FIL0}:$L${FIL1},"Por debajo del objetivo")', fmt=ENT)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Comparativa
# --------------------------------------------------------------------------
DIAGNOSTICO = (
    'IF($H{r}=0,'
    '"Las cuatro lecturas coinciden en lo mejor: mantener, destacar en carta y '
    'proteger receta y proveedor.",'
    'IF(AND($D{r}="Dog",$E{r}="Loser",$F{r}="Problem"),'
    '"Poco pedido, food cost alto y aporta poco margen al total: candidato a '
    'retirar o a reformular de raíz.",'
    'IF(AND($D{r}="Plowhorse",$E{r}="Winner"),'
    '"Popular con margen bajo: subir precio o bajar coste.",'
    'IF(AND($D{r}="Plowhorse",$E{r}="Loser"),'
    '"Popular pero caro de producir y de margen corto: reformular la receta '
    'antes de tocar el precio.",'
    'IF(AND(' + DAT + '$G{r}>=' + DAT + '$N{r},' + DAT + '$H{r}>' + DAT +
    '$O{r}),'
    '"MC alto con food cost pobre: proteger el margen en euros, revisar precio, '
    'no retirar.",'
    'IF(AND($D{r}="Puzzle",$E{r}="Winner"),'
    '"Rentable y con food cost sano, pero poco pedido: es problema de carta, no '
    'de cocina. Renómbralo, súbelo de posición o que lo sugiera la sala.",'
    'IF($F{r}="Sleeper",'
    '"Food cost sano pero aporta poco margen al total: dale visibilidad antes '
    'de retirarlo.",'
    'IF(AND($G{r}="Por debajo del objetivo",$D{r}="Star"),'
    '"Vende y deja margen, pero se queda por debajo del Goal Value de su '
    'familia: va justo de recorrido, revisa el precio.",'
    'IF($D{r}="Dog",'
    '"Poco pedido: antes de retirarlo comprueba si es de temporada o si '
    'acompaña a otro plato que sí vende.",'
    '"Las lecturas discrepan: cada método mide dos de las tres variables '
    '(popularidad, margen en euros y food cost). Decide con la que más te duela '
    'este mes.")))))))))')

DECISION = (
    'IF($H{r}=0,"Mantener",'
    'IF(AND($D{r}="Dog",$E{r}="Loser"),"Retirar",'
    'IF($D{r}="Plowhorse","Resubir",'
    'IF($D{r}="Puzzle","Rediseñar",'
    'IF($F{r}="Problem","Reformular","Revisar")))))')


def hoja_comparativa(wb):
    ws = wb.create_sheet('Comparativa')
    cabecera(ws, 'Comparativa — dónde discrepan los cuatro métodos')
    apunte(ws, 'E3', 'Coincidir no da más seguridad: sólo dice que el plato se '
                     've igual desde cualquier ángulo. El valor está en las '
                     'filas que discrepan.')
    encabezados(ws, 4, [
        ('A', '#', 5), ('B', 'Plato', 44), ('C', 'Familia', 14),
        ('D', 'Kasavana & Smith', 15), ('E', 'Miller', 13),
        ('F', 'Pavesic', 13), ('G', 'Goal Value', 22),
        ('H', 'Lecturas fuera de la mejor categoría', 14),
        ('I', 'Diagnóstico', 78), ('J', 'Decisión sugerida', 16),
    ])
    ws.freeze_panes = 'B5'
    for i in range(FIL1 - FIL0 + 1):
        r = FIL0 + i
        motor.val(ws, f'A{r}', i + 1, fmt=ENT)
        espejo(ws, 'B', r, 'B')
        espejo(ws, 'C', r, 'C')
        motor.f(ws, f'D{r}', f'=IFERROR(IF({KS}$K{r}="","",{KS}$K{r}),"")',
                bold=True)
        motor.f(ws, f'E{r}', f'=IFERROR(IF({MIL}$K{r}="","",{MIL}$K{r}),"")',
                bold=True)
        motor.f(ws, f'F{r}', f'=IFERROR(IF({PAV}$K{r}="","",{PAV}$K{r}),"")',
                bold=True)
        motor.f(ws, f'G{r}', f'=IFERROR(IF({GVA}$L{r}="","",{GVA}$L{r}),"")',
                bold=True)
        motor.f(ws, f'H{r}',
                f'=IFERROR(IF($B{r}="","",'
                f'IF($D{r}="Star",0,1)+IF($E{r}="Winner",0,1)+'
                f'IF($F{r}="Prime",0,1)+'
                f'IF($G{r}="Por encima del objetivo",0,1)),"")', fmt=ENT)
        ws[f'H{r}'].font = Font(bold=True, color=AZUL)
        motor.f(ws, f'I{r}',
                f'=IFERROR(IF($B{r}="","",' + DIAGNOSTICO.format(r=r) + '),"")')
        ws[f'I{r}'].alignment = Alignment(vertical='top', wrap_text=True)
        motor.f(ws, f'J{r}',
                f'=IFERROR(IF($B{r}="","",' + DECISION.format(r=r) + '),"")',
                bold=True)
    cf_reglas(ws, f'H{FIL0}:H{FIL1}', [
        (f'=AND(ISNUMBER($H{FIL0}),$H{FIL0}=0)',
         motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        (f'=AND(ISNUMBER($H{FIL0}),$H{FIL0}>=3)',
         motor.CF_ROJO_BG, motor.CF_ROJO_FG),
    ])

    bloque(ws, 'A31', 'RESUMEN — lo calcula el libro')
    encabezados(ws, 32, [
        ('B', 'Lecturas fuera de la mejor categoría', None),
        ('C', 'Platos', None), ('D', '% de la carta', None),
    ], alto=34)
    for k in range(5):
        r = 33 + k
        motor.val(ws, f'B{r}', k, fmt=ENT, bold=True)
        motor.f(ws, f'C{r}', f'=COUNTIF($H${FIL0}:$H${FIL1},$B{r})', fmt=ENT)
        motor.f(ws, f'D{r}',
                f'=IFERROR(IF(COUNTIF($B${FIL0}:$B${FIL1},"<>")=0,"",'
                f'$C{r}/COUNTIF($B${FIL0}:$B${FIL1},"<>")),"")', fmt=PCT)
    motor.val(ws, 'B39', 'Platos con las cuatro lecturas en la mejor categoría',
              bold=True)
    motor.f(ws, 'C39', f'=COUNTIF($H${FIL0}:$H${FIL1},0)', fmt=ENT)
    motor.val(ws, 'B40', 'Platos con tres o cuatro lecturas fuera', bold=True)
    motor.f(ws, 'C40',
            f'=COUNTIF($H${FIL0}:$H${FIL1},3)+COUNTIF($H${FIL0}:$H${FIL1},4)',
            fmt=ENT)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Menú Precio Fijo»
# --------------------------------------------------------------------------
MF0, MF1 = 5, 16
CURSOS = ['Primeros', 'Segundos', 'Postres']
ANCHOS_MENU = [('A', 34), ('B', 34), ('C', 16), ('D', 16), ('E', 17),
               ('F', 17), ('G', 52)]
# Escenario A = el comensal se va a la opción cara; B = a la barata.
MIX_A = {'Primeros': [0.25, 0.20, 0.55], 'Segundos': [0.30, 0.45, 0.25],
         'Postres': [0.40, 0.60]}
MIX_B = {'Primeros': [0.35, 0.45, 0.20], 'Segundos': [0.60, 0.15, 0.25],
         'Postres': [0.70, 0.30]}


def hoja_menu(wb):
    ws = wb.create_sheet('Menú Precio Fijo')
    for letra, ancho in ANCHOS_MENU:
        ws.column_dimensions[letra].width = ancho
    cabecera(ws, 'Menú de precio fijo — el margen lo decide el mix')
    apunte(ws, 'D3', 'En un menú cerrado el precio ya está puesto: lo que mueve '
                     'el margen es qué elige la gente.')
    encabezados(ws, 4, [
        ('A', 'Curso', None), ('B', 'Opción', None),
        ('C', 'Coste por ración (€)', None), ('D', 'Mix base (%)', None),
        ('E', 'Mix escenario A (%)', None), ('F', 'Mix escenario B (%)', None),
        ('G', 'Nota', None),
    ], alto=34)
    ws.freeze_panes = 'A5'

    filas, r = [], MF0
    for curso, opciones in DE.MENU_PRECIO_FIJO['cursos']:
        for k, (nombre, coste, mix) in enumerate(opciones):
            motor.val(ws, f'A{r}', curso)
            motor.val(ws, f'B{r}', nombre)
            motor.val(ws, f'C{r}', coste, fmt=EUR)
            motor.val(ws, f'D{r}', mix, fmt=PCT0)
            motor.val(ws, f'E{r}', MIX_A[curso][k], fmt=PCT0)
            motor.val(ws, f'F{r}', MIX_B[curso][k], fmt=PCT0)
            filas.append(r)
            r += 1
        motor.val(ws, f'A{r}', curso)              # fila libre de ese curso
        ws[f'C{r}'].number_format = EUR
        for col in 'DEF':
            ws[f'{col}{r}'].number_format = PCT0
        filas.append(r)
        r += 1
    while r <= MF1:                                # colchón final
        ws[f'C{r}'].number_format = EUR
        for col in 'DEF':
            ws[f'{col}{r}'].number_format = PCT0
        filas.append(r)
        r += 1
    for fila in filas:
        motor.verde(ws, f'A{fila}:F{fila}')
    motor.dv_lista(ws, [f'A{f}' for f in filas], CURSOS,
                   titulo='Curso no válido')
    motor.dv_numerica(ws, [f'C{f}' for f in filas], minimo=0)
    motor.dv_porcentaje(ws, [f'{c}{f}' for f in filas for c in 'DEF'],
                        titulo='Mix del curso',
                        prompt='Porcentaje de comensales que eligen esta '
                               'opción. El mix de cada curso tiene que sumar '
                               '100' + N + '%.')
    motor.val(ws, 'G5', 'Cada curso suma 100' + N + '%: es el reparto de '
                        'elecciones, no un descuento.')
    motor.val(ws, 'G9', 'Escenario A: el comensal se va a la opción más cara. '
                        'Escenario B: se va a la más barata.')

    bloque(ws, 'A18', 'PARÁMETROS DEL MENÚ')
    p_pvp = verde_propio(
        ws, 19, 'A', 'C', 'PVP del menú, con IVA (€)',
        DE.MENU_PRECIO_FIJO['pvp_con_iva'], EUR,
        'El precio que ve el comensal en la pizarra o en la carta.',
        col_nota='E')
    p_iva = parametro(ws, 20, 'A', 'C', 'iva_restauracion', col_nota='E')
    motor.dv_porcentaje(ws, ['C20'], titulo='Tipo de IVA',
                        prompt='Se escribe en tanto por uno: 0,10 = 10' + N +
                               '%.')
    motor.val(ws, 'A21', 'PVP del menú, sin IVA (€)', bold=True)
    motor.f(ws, 'C21',
            f'=IFERROR(IF(OR({p_pvp}="",{p_iva}=""),"",{p_pvp}/(1+{p_iva})),"")',
            fmt=EUR, bold=True)
    p_fijos = verde_propio(
        ws, 22, 'A', 'C', 'Costes fijos por menú (€)',
        DE.MENU_PRECIO_FIJO['fijos_por_menu'], EUR,
        'Pan, agua o copa de vino de la casa y café si van incluidos, '
        'prorrateados por menú servido.', col_nota='E')
    p_obj = verde_propio(
        ws, 23, 'A', 'C', 'Food cost objetivo del menú (%)', 0.30, PCT,
        'Objetivo de la casa. Si el menú se pasa de aquí, la fila «Food cost '
        'del menú» se pone en rojo.', col_nota='E')
    motor.dv_porcentaje(ws, ['C23'], titulo='Food cost objetivo',
                        prompt='Se escribe en tanto por uno: 0,30 = 30' + N +
                               '%.')
    p_menus = verde_propio(
        ws, 24, 'A', 'C', 'Menús servidos al mes', 100, ENT,
        'Cuántos menús de este tipo sirves al mes: escala el margen por menú '
        'al margen del mes.', col_nota='E')
    motor.dv_numerica(ws, ['C24'], minimo=0, titulo='Menús servidos',
                      mensaje='Escribe un número entero de menús servidos al '
                              'mes (0 o más).')

    bloque(ws, 'A25', 'COSTE MEDIO PONDERADO POR CURSO — lo calcula el libro')
    encabezados(ws, 26, [
        ('A', 'Curso', None), ('B', 'Suma del mix base (%)', None),
        ('C', 'Coste medio base (€)', None), ('D', 'Suma del mix A (%)', None),
        ('E', 'Coste medio A (€)', None), ('F', 'Suma del mix B (%)', None),
        ('G', 'Coste medio B (€)', None),
    ], alto=34)
    for k, curso in enumerate(CURSOS):
        r = 27 + k
        motor.val(ws, f'A{r}', curso, bold=True)
        sel = f'--($A${MF0}:$A${MF1}=$A{r})'
        for col_mix, col_suma, col_coste in (('D', 'B', 'C'), ('E', 'D', 'E'),
                                             ('F', 'F', 'G')):
            motor.f(ws, f'{col_suma}{r}',
                    f'=IFERROR(SUMIF($A${MF0}:$A${MF1},$A{r},'
                    f'${col_mix}${MF0}:${col_mix}${MF1}),"")', fmt=PCT0)
            motor.f(ws, f'{col_coste}{r}',
                    f'=IFERROR(SUMPRODUCT({sel},$C${MF0}:$C${MF1},'
                    f'${col_mix}${MF0}:${col_mix}${MF1}),"")', fmt=EUR)
    for col in ('B', 'D', 'F'):
        motor.regla_expresion(
            ws, f'{col}27:{col}29',
            f'=AND(ISNUMBER({col}27),ROUND({col}27,4)<>1)')
    motor.val(ws, 'A30', 'Si una suma de mix no da 100' + N + '% se pone en '
                         'rojo: el reparto de ese curso está mal repartido.')

    bloque(ws, 'A32', 'RESULTADO DEL MENÚ — lo calcula el libro')
    encabezados(ws, 33, [
        ('A', 'Concepto', None), ('B', 'Mix base', None),
        ('C', 'Escenario A', None), ('D', 'Escenario B', None),
    ], alto=24)
    for k, etiqueta in enumerate(['Coste medio de los primeros (€)',
                                  'Coste medio de los segundos (€)',
                                  'Coste medio de los postres (€)',
                                  'Costes fijos por menú (€)']):
        motor.val(ws, f'A{34 + k}', etiqueta)
    for col, origen in (('B', 'C'), ('C', 'E'), ('D', 'G')):
        motor.f(ws, f'{col}34', f'=IFERROR(${origen}$27,"")', fmt=EUR)
        motor.f(ws, f'{col}35', f'=IFERROR(${origen}$28,"")', fmt=EUR)
        motor.f(ws, f'{col}36', f'=IFERROR(${origen}$29,"")', fmt=EUR)
        motor.f(ws, f'{col}37', f'=IFERROR({p_fijos},"")', fmt=EUR)
        total(ws, f'{col}38', f'=IFERROR(SUM({col}34:{col}37),"")', fmt=EUR,
              formula=True)
        motor.f(ws, f'{col}39',
                f'=IFERROR(IF(OR($C$21="",$C$21=0,{col}38=""),"",'
                f'{col}38/$C$21),"")', fmt=PCT, bold=True)
        motor.f(ws, f'{col}40',
                f'=IFERROR(IF(OR($C$21="",{col}38=""),"",$C$21-{col}38),"")',
                fmt=EUR, bold=True)
        motor.f(ws, f'{col}41',
                f'=IFERROR(IF({col}40="","",{col}40*{p_menus}),"")', fmt=EUR)
    total(ws, 'A38', 'Coste medio total del menú (€)')
    motor.val(ws, 'A39', 'Food cost del menú (%)', bold=True)
    motor.val(ws, 'A40', 'Margen de contribución por menú (€)', bold=True)
    motor.val(ws, 'A41', 'Margen sobre los menús servidos al mes (€)')
    motor.regla_expresion(ws, 'B39:D39', f'=AND(ISNUMBER(B$39),B$39>{p_obj})')
    motor.val(ws, 'F39', 'Rojo = por encima del food cost objetivo del menú.')
    motor.val(ws, 'A43',
              'Los tres escenarios usan el MISMO precio y los MISMOS costes: lo '
              'único que cambia es el reparto de elecciones. Esa diferencia es '
              'el margen que te juegas cada día en el menú.')
    motor.val(ws, 'A44',
              'En el ejemplo el menú se queda POR ENCIMA del food cost objetivo '
              'de la casa. Es lo habitual en un menú de precio cerrado: o subes '
              'el precio, o cambias las opciones que ofreces, o asumes que el '
              'menú trae clientela y el margen lo hace la carta. Lo que no vale '
              'es no saberlo.')
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Mapa de celdas (contrato con el guion de la guía)
# --------------------------------------------------------------------------
COLS_MAPA_DATOS = [[c[1], c[0], t] for c, t in zip(COLS_DATOS, [
    'num', 'txt', 'txt', 'num', 'eur', 'eur', 'eur', 'pct1', 'eur', 'num',
    'num', 'pct1', 'pct1', 'eur', 'pct1', 'eur', 'eur'])]


def mapa():
    return {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Datos': {
                'celdas': {
                    'Total de unidades vendidas de la carta': 'D32',
                    'Margen de contribución medio ponderado de la carta': 'G32',
                    'Food cost medio ponderado de la carta': 'H32',
                    'Ventas netas del mes de la carta': 'I32',
                    'Factor del umbral de popularidad': 'D33',
                    'Tipo de IVA de restauración en sala': 'D34',
                    'Platos con ventas en la carta': 'D35',
                    'Platos dados de alta en la carta': 'D36',
                    'Entrantes: platos en carta': 'C40',
                    'Entrantes: uds vendidas': 'E40',
                    'Entrantes: mix sobre la carta': 'F40',
                    'Entrantes: MC medio ponderado': 'G40',
                    'Entrantes: food cost medio ponderado': 'H40',
                    'Entrantes: ventas netas del mes': 'I40',
                    'Principales: platos en carta': 'C41',
                    'Principales: uds vendidas': 'E41',
                    'Principales: mix sobre la carta': 'F41',
                    'Principales: MC medio ponderado': 'G41',
                    'Principales: food cost medio ponderado': 'H41',
                    'Principales: ventas netas del mes': 'I41',
                    'Postres: platos en carta': 'C42',
                    'Postres: uds vendidas': 'E42',
                    'Postres: mix sobre la carta': 'F42',
                    'Postres: MC medio ponderado': 'G42',
                    'Postres: food cost medio ponderado': 'H42',
                    'Postres: ventas netas del mes': 'I42',
                    'Menú: platos en carta': 'C43',
                    'Menú: uds vendidas': 'E43',
                },
                'tablas': [
                    {'titulo': 'Carta: 20 platos de ejemplo y 5 filas libres',
                     'cols': COLS_MAPA_DATOS, 'filas': [5, 29]},
                    {'titulo': 'Resumen por familia',
                     'cols': [['Familia', 'B', 'txt'],
                              ['Platos en carta', 'C', 'num'],
                              ['Platos con ventas', 'D', 'num'],
                              ['Uds vendidas', 'E', 'num'],
                              ['Mix sobre la carta (%)', 'F', 'pct1'],
                              ['MC medio ponderado (€)', 'G', 'eur'],
                              ['Food cost medio ponderado (%)', 'H', 'pct1'],
                              ['Ventas netas del mes (€)', 'I', 'eur']],
                     'filas': [40, 43]},
                ],
            },
            'Kasavana-Smith': {
                'celdas': {
                    'Platos Star': 'C33', 'Uds de los Star': 'D33',
                    '% de uds de los Star': 'E33',
                    'MC total aportado por los Star': 'F33',
                    'Platos Plowhorse': 'C34', 'Uds de los Plowhorse': 'D34',
                    'MC total aportado por los Plowhorse': 'F34',
                    'Platos Puzzle': 'C35',
                    'MC total aportado por los Puzzle': 'F35',
                    'Platos Dog': 'C36', 'Uds de los Dog': 'D36',
                    'MC total aportado por los Dog': 'F36',
                },
                'tablas': [
                    {'titulo': 'Clasificación Kasavana & Smith plato a plato',
                     'cols': [['#', 'A', 'num'], ['Plato', 'B', 'txt'],
                              ['Familia', 'C', 'txt'],
                              ['Uds vendidas', 'D', 'num'],
                              ['Mix dentro de su familia (%)', 'E', 'pct1'],
                              ['Umbral de popularidad de su familia (%)', 'F',
                               'pct1'],
                              ['Popularidad', 'G', 'txt'],
                              ['MC del plato (€)', 'H', 'eur'],
                              ['MC medio ponderado de su familia (€)', 'I',
                               'eur'],
                              ['Nivel de margen', 'J', 'txt'],
                              ['Clasificación', 'K', 'txt'],
                              ['Qué significa', 'L', 'txt'],
                              ['Acción recomendada', 'M', 'txt']],
                     'filas': [5, 29]},
                    {'titulo': 'Resumen por clasificación',
                     'cols': [['Clasificación', 'B', 'txt'],
                              ['Platos', 'C', 'num'],
                              ['Uds vendidas', 'D', 'num'],
                              ['% de las uds de la carta', 'E', 'pct1'],
                              ['MC total aportado (€)', 'F', 'eur']],
                     'filas': [33, 36]},
                ],
            },
            'Miller': {
                'celdas': {
                    'Platos Winner': 'C33', 'Uds de los Winner': 'D33',
                    '% de uds de los Winner': 'E33',
                    'Food cost medio ponderado de los Winner': 'F33',
                    'Platos Marginal': 'C34',
                    'Food cost medio ponderado de los Marginal': 'F34',
                    'Platos Loser': 'C35', '% de uds de los Loser': 'E35',
                    'Food cost medio ponderado de los Loser': 'F35',
                },
                'tablas': [
                    {'titulo': 'Clasificación Miller plato a plato',
                     'cols': [['#', 'A', 'num'], ['Plato', 'B', 'txt'],
                              ['Familia', 'C', 'txt'],
                              ['Uds vendidas', 'D', 'num'],
                              ['Mix dentro de su familia (%)', 'E', 'pct1'],
                              ['Umbral de popularidad de su familia (%)', 'F',
                               'pct1'],
                              ['Popularidad', 'G', 'txt'],
                              ['Food cost del plato (%)', 'H', 'pct1'],
                              ['Food cost medio ponderado de su familia (%)',
                               'I', 'pct1'],
                              ['Nivel de food cost', 'J', 'txt'],
                              ['Clasificación', 'K', 'txt'],
                              ['Qué significa', 'L', 'txt'],
                              ['Acción recomendada', 'M', 'txt']],
                     'filas': [5, 29]},
                    {'titulo': 'Resumen por clasificación',
                     'cols': [['Clasificación', 'B', 'txt'],
                              ['Platos', 'C', 'num'],
                              ['Uds vendidas', 'D', 'num'],
                              ['% de las uds de la carta', 'E', 'pct1'],
                              ['Food cost medio ponderado del grupo (%)', 'F',
                               'pct1']],
                     'filas': [33, 35]},
                ],
            },
            'Pavesic': {
                'celdas': {
                    'Platos Prime': 'C33', 'Uds de los Prime': 'D33',
                    '% de uds de los Prime': 'E33',
                    'MC ponderado total de los Prime': 'F33',
                    'Platos Standard': 'C34',
                    'MC ponderado total de los Standard': 'F34',
                    'Platos Sleeper': 'C35',
                    'MC ponderado total de los Sleeper': 'F35',
                    'Platos Problem': 'C36', '% de uds de los Problem': 'E36',
                    'MC ponderado total de los Problem': 'F36',
                },
                'tablas': [
                    {'titulo': 'Clasificación Pavesic plato a plato',
                     'cols': [['#', 'A', 'num'], ['Plato', 'B', 'txt'],
                              ['Familia', 'C', 'txt'],
                              ['Uds vendidas', 'D', 'num'],
                              ['Food cost del plato (%)', 'E', 'pct1'],
                              ['Food cost medio ponderado de su familia (%)',
                               'F', 'pct1'],
                              ['Nivel de food cost', 'G', 'txt'],
                              ['MC ponderado: MC × uds (€)', 'H', 'eur'],
                              ['Media del MC ponderado de su familia (€)', 'I',
                               'eur'],
                              ['Nivel de MC ponderado', 'J', 'txt'],
                              ['Clasificación', 'K', 'txt'],
                              ['Qué significa', 'L', 'txt'],
                              ['Acción recomendada', 'M', 'txt']],
                     'filas': [5, 29]},
                    {'titulo': 'Resumen por clasificación',
                     'cols': [['Clasificación', 'B', 'txt'],
                              ['Platos', 'C', 'num'],
                              ['Uds vendidas', 'D', 'num'],
                              ['% de las uds de la carta', 'E', 'pct1'],
                              ['MC ponderado total del grupo (€)', 'F', 'eur']],
                     'filas': [33, 36]},
                ],
            },
            'Goal Value': {
                'celdas': {
                    'Coste de personal sobre ventas': 'D33',
                    'Otros costes variables sobre ventas': 'D34',
                    'Goal Value medio de la carta': 'D35',
                    'Platos por encima del objetivo de su familia': 'D36',
                    'Platos por debajo del objetivo de su familia': 'D37',
                },
                'tablas': [
                    {'titulo': 'Goal Value plato a plato',
                     'cols': [['#', 'A', 'num'], ['Plato', 'B', 'txt'],
                              ['Familia', 'C', 'txt'],
                              ['Uds vendidas', 'D', 'num'],
                              ['Food cost del plato (%)', 'E', 'pct1'],
                              ['PVP sin IVA (€)', 'F', 'eur'],
                              ['Goal Value del plato', 'G', 'num2'],
                              ['Uds medias de su familia', 'H', 'num'],
                              ['PVP medio ponderado de su familia (€)', 'I',
                               'eur'],
                              ['Food cost medio ponderado de su familia (%)',
                               'J', 'pct1'],
                              ['Goal Value objetivo de su familia', 'K', 'num2'],
                              ['Lectura', 'L', 'txt'],
                              ['Distancia al objetivo (índice)', 'M', 'num2'],
                              ['Qué revisar', 'N', 'txt']],
                     'filas': [5, 29]},
                ],
            },
            'Comparativa': {
                'celdas': {
                    'Platos con 0 lecturas fuera de la mejor categoría': 'C33',
                    'Platos con 1 lectura fuera': 'C34',
                    'Platos con 2 lecturas fuera': 'C35',
                    'Platos con 3 lecturas fuera': 'C36',
                    'Platos con 4 lecturas fuera': 'C37',
                    '% de la carta con 0 lecturas fuera': 'D33',
                    '% de la carta con 2 lecturas fuera': 'D35',
                    'Platos con las cuatro lecturas en la mejor categoría':
                        'C39',
                    'Platos con tres o cuatro lecturas fuera': 'C40',
                },
                'tablas': [
                    {'titulo': 'Las cuatro lecturas del mismo plato',
                     'cols': [['#', 'A', 'num'], ['Plato', 'B', 'txt'],
                              ['Familia', 'C', 'txt'],
                              ['Kasavana & Smith', 'D', 'txt'],
                              ['Miller', 'E', 'txt'], ['Pavesic', 'F', 'txt'],
                              ['Goal Value', 'G', 'txt'],
                              ['Lecturas fuera de la mejor categoría', 'H',
                               'num'],
                              ['Diagnóstico', 'I', 'txt'],
                              ['Decisión sugerida', 'J', 'txt']],
                     'filas': [5, 29]},
                    {'titulo': 'Cuántos platos discrepan',
                     'cols': [['Lecturas fuera de la mejor categoría', 'B',
                               'num'],
                              ['Platos', 'C', 'num'],
                              ['% de la carta', 'D', 'pct1']],
                     'filas': [33, 37]},
                ],
            },
            'Menú Precio Fijo': {
                'celdas': {
                    'PVP del menú con IVA': 'C19',
                    'Tipo de IVA de restauración en sala': 'C20',
                    'PVP del menú sin IVA': 'C21',
                    'Costes fijos por menú': 'C22',
                    'Food cost objetivo del menú': 'C23',
                    'Menús servidos al mes': 'C24',
                    'Coste medio de los primeros (mix base)': 'C27',
                    'Coste medio de los segundos (mix base)': 'C28',
                    'Coste medio de los postres (mix base)': 'C29',
                    'Coste medio de los primeros (escenario A)': 'E27',
                    'Coste medio de los primeros (escenario B)': 'G27',
                    'Coste medio total del menú (mix base)': 'B38',
                    'Coste medio total del menú (escenario A)': 'C38',
                    'Coste medio total del menú (escenario B)': 'D38',
                    'Food cost del menú (mix base)': 'B39',
                    'Food cost del menú (escenario A)': 'C39',
                    'Food cost del menú (escenario B)': 'D39',
                    'Margen de contribución por menú (mix base)': 'B40',
                    'Margen de contribución por menú (escenario A)': 'C40',
                    'Margen de contribución por menú (escenario B)': 'D40',
                    'Margen sobre los menús servidos al mes (mix base)': 'B41',
                    'Margen sobre los menús servidos al mes (escenario A)': 'C41',
                    'Margen sobre los menús servidos al mes (escenario B)': 'D41',
                },
                'tablas': [
                    {'titulo': 'Opciones del menú por curso',
                     'cols': [['Curso', 'A', 'txt'], ['Opción', 'B', 'txt'],
                              ['Coste por ración (€)', 'C', 'eur'],
                              ['Mix base (%)', 'D', 'pct0'],
                              ['Mix escenario A (%)', 'E', 'pct0'],
                              ['Mix escenario B (%)', 'F', 'pct0']],
                     'filas': [5, 16]},
                    {'titulo': 'Coste medio ponderado por curso',
                     'cols': [['Curso', 'A', 'txt'],
                              ['Suma del mix base (%)', 'B', 'pct0'],
                              ['Coste medio base (€)', 'C', 'eur'],
                              ['Suma del mix A (%)', 'D', 'pct0'],
                              ['Coste medio A (€)', 'E', 'eur'],
                              ['Suma del mix B (%)', 'F', 'pct0'],
                              ['Coste medio B (€)', 'G', 'eur']],
                     'filas': [27, 29]},
                    {'titulo': 'Resultado del menú por escenario de mix',
                     'cols': [['Concepto', 'A', 'txt'],
                              ['Mix base', 'B', 'eur'],
                              ['Escenario A', 'C', 'eur'],
                              ['Escenario B', 'D', 'eur']],
                     'filas': [34, 41]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_datos(wb)
    hoja_kasavana(wb)
    hoja_miller(wb)
    hoja_pavesic(wb)
    hoja_goal_value(wb)
    hoja_comparativa(wb)
    hoja_menu(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-22s %d' % (hoja, n))


if __name__ == '__main__':
    main()
