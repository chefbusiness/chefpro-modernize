#!/usr/bin/env python3
"""
gen_cuadro-de-mando-prime-cost.py — libro 7 de «Guía Food Cost + Ingeniería de
Menú» (SPEC §2.2 fila 7, decisión D5).

Hojas: Instrucciones · Parámetros · Mensual (con gráfico de líneas).

Para qué sirve: ver el coste de producto y el de personal JUNTOS, mes a mes.
Un food cost del 28 % no dice nada si el personal se te va al 40 %: lo que mide
la salud del negocio es la suma de los dos, el prime cost.

Umbral español (D5): producto ~30 % + personal 30-35 % con servicio en mesa
(15-25 % en barra o autoservicio) según CaixaBankLab × elBulliFoundation →
objetivo ≤ 65 % con servicio en mesa y ≤ 55 % en barra o autoservicio. El 60 %
de Toast se cita como contraste de EE. UU., no como objetivo.

Convenciones de la familia (motor.py de guias-v2_0): celdas verdes = entrada;
ninguna celda con fórmula en verde; cero constantes dentro de una fórmula (la
Seguridad Social y los objetivos viven en celdas); «sin dato» = "" nunca 0;
IFERROR en todo cociente; semáforos con ISNUMBER; prohibidas INDIRECT, COUNTA,
PMT, OFFSET, XLOOKUP, LET, LAMBDA y matrices dinámicas.

El objetivo activo y el food cost objetivo se ESPEJAN en la hoja «Mensual»
(filas 20-22): el formato condicional no puede referirse a otra hoja en Google
Sheets, y la única manera de hacerlo sería INDIRECT, que está prohibida.

Salida: build/cuadro-de-mando-prime-cost.xlsx
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
sys.path.insert(0, AQUI)

import motor                                                   # noqa: E402
import datos_ejemplo as D                                      # noqa: E402

motor.CTX['producto'] = 'guia-food-cost-ingenieria-menu'

PRODUCTO = 'Guía Food Cost + Ingeniería de Menú'
SUBTITULO = 'AI Chef Pro · aichef.pro — ' + PRODUCTO
VERSION = ('Versión 1.0 · septiembre 2026 · '
           'aichef.pro/guia-food-cost-ingenieria-menu · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 '
       '· johnguerrero.es')
NOTA_DESPROTEGER = ('Para editar la estructura o una celda que no esté en '
                    'verde, desprotege la hoja (sin contraseña).')
TITULO = 'Cuadro de mando del prime cost'
NOMBRE = 'cuadro-de-mando-prime-cost'

FMT_EUR = '#,##0.00 €'
FMT_PCT = '0.0%'

GRIS = 'F2F2F2'
CREMA = 'FFF6DC'
ORO = 'FFD700'
CABECERA = '2D2D2D'

#: URL de la fuente del umbral español, extraída del propio juego de datos.
URL_CAIXA = D.FUENTE_BEVERAGE[D.FUENTE_BEVERAGE.index('(') + 1:
                              D.FUENTE_BEVERAGE.index(')')]
TIPOS_NEGOCIO = list(D.PRIME_COST_OBJETIVO.keys())

# Filas de la hoja Parámetros
P_TIPO = 5
P_CAB = 7
P_INI = 8
P_FIN = P_INI + len(TIPOS_NEGOCIO) - 1
P_ACTIVO = 11
P_SS = 20
P_FC = 21

# Filas de la hoja Mensual
M_CAB, M_INI = 4, 5
M_FIN = M_INI + len(D.MESES) - 1
M_TOT = M_FIN + 1
M_ESP = M_TOT + 3            # espejo de parámetros en la propia hoja


# --------------------------------------------------------------------------
def cabecera(ws, fila, columnas, altura=44):
    for letra, texto in columnas:
        c = ws[letra + str(fila)]
        c.value = texto
        c.fill = PatternFill('solid', fgColor=CABECERA)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    ws.row_dimensions[fila].height = altura


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12)


def fila_total(ws, fila, primera, ultima):
    for col in range(motor.column_index_from_string(primera),
                     motor.column_index_from_string(ultima) + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = PatternFill('solid', fgColor=CREMA)
        c.font = Font(bold=True)


def cf_expresion(ws, rango, formula, bg, fg):
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


def pagina(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezar(ws, titulo, nota=None):
    motor.val(ws, 'A1', titulo)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    if nota:
        motor.val(ws, 'A3', nota)
        ws['A3'].font = Font(italic=True, size=9)


def anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


# --------------------------------------------------------------------------
PASOS = [
    '1. Hoja «Parámetros»: elige tu tipo de negocio. El objetivo de prime cost '
    'cambia solo (65 % con servicio en mesa, 55 % en barra o autoservicio) y '
    'los dos objetivos son celdas verdes que puedes ajustar.',
    '2. Revisa la Seguridad Social a cargo de la empresa y tu food cost '
    'objetivo: el libro entero lee esas dos celdas.',
    '3. Hoja «Mensual»: por cada mes, ventas netas de comida y de bebida, '
    'stock inicial, compras, stock final, salarios brutos y otros costes de '
    'personal. Todo SIN IVA.',
    '4. El consumo de materia prima se calcula solo: stock inicial + compras − '
    'stock final. No es lo mismo que las compras del mes.',
    '5. El coste de personal añade la Seguridad Social a los brutos: un '
    'salario de 15.400 € cuesta 20.482 € con un 33 % de cotización. Contar '
    'sólo los brutos es el error más caro de esta hoja.',
    '6. Lee la columna «Prime cost»: si está por encima del objetivo, se pone '
    'en rojo. El gráfico de abajo enseña los doce meses contra la línea del '
    'objetivo.',
    '7. La fila TOTAL/MEDIA es una media PONDERADA: divide el consumo del año '
    'entre las ventas del año, no hace la media de los doce porcentajes.',
]

NOTAS_LIBRO = [
    'Todas las cifras van SIN IVA. El food cost y el prime cost se miden sobre '
    'la venta NETA (base imponible) y con el coste NETO de IVA soportado: el '
    'IVA de las compras se deduce en el modelo 303, es tesorería, no coste.',
    'Prime cost = coste de materia prima + coste de personal con Seguridad '
    'Social, sobre las ventas netas. Es la métrica que mide la salud del '
    'negocio porque junta las dos partidas que de verdad puedes mover.',
    'Un food cost «bueno» puede estar tapando un labor cost roto: por eso las '
    'dos columnas van juntas y el semáforo mira la suma.',
    'El margen tras prime cost no es el beneficio: todavía no ha pagado '
    'alquiler, suministros, amortizaciones ni impuestos.',
    'La columna «Coste de materia prima sobre ventas totales» de esta hoja NO '
    'es el mismo número que el «food cost' + motor.NARROW + '%» de la carta '
    '(«matriz-multimetodo-carta.xlsx»): aquí se compara el consumo de TODA la '
    'materia prima (comida + bebida) contra las ventas TOTALES, porque es lo '
    'que necesita el prime cost. La matriz calcula el food cost SOLO de la '
    'comida, sobre las ventas SOLO de comida. Los dos son correctos; miden '
    'cosas distintas.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    anchos(ws, {'A': 110.0})
    motor.val(ws, 'A1', TITULO)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    seccion(ws, 'A4', 'Instrucciones de uso')
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), 'Celdas verdes = campos editables')
    ws['A' + str(fila)].fill = PatternFill('solid', fgColor=motor.VERDE)
    fila += 2
    for nota in NOTAS_LIBRO:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), NOTA_DESPROTEGER)
    fila += 1
    motor.val(ws, 'A' + str(fila), BIO)
    fila += 1
    motor.val(ws, 'A' + str(fila), VERSION)
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
def hoja_parametros(wb):
    ws = wb.create_sheet('Parámetros')
    anchos(ws, {'A': 42, 'B': 20, 'C': 18, 'D': 18})
    encabezar(ws, 'Parámetros del cuadro de mando')

    seccion(ws, 'A4', 'TIPO DE NEGOCIO Y OBJETIVO DE PRIME COST')
    motor.val(ws, 'A%d' % P_TIPO, 'Tipo de negocio', bold=True)
    motor.val(ws, 'B%d' % P_TIPO, TIPOS_NEGOCIO[0], verde_=True)
    cabecera(ws, P_CAB, [('A', 'Tipo de negocio'),
                         ('B', 'Objetivo de prime cost (%)')], altura=32)
    for i, tipo in enumerate(TIPOS_NEGOCIO):
        fila = P_INI + i
        motor.val(ws, 'A%d' % fila, tipo, bold=True)
        motor.val(ws, 'B%d' % fila, D.PRIME_COST_OBJETIVO[tipo], fmt=FMT_PCT,
                  verde_=True)
    motor.val(ws, 'A%d' % P_ACTIVO, 'Objetivo de prime cost en vigor (%)',
              bold=True)
    motor.f(ws, 'B%d' % P_ACTIVO,
            '=IF($B${t}=$A${a},$B${a},IF($B${t}=$A${b},$B${b},""))'
            .format(t=P_TIPO, a=P_INI, b=P_FIN), fmt=FMT_PCT)
    ws['B%d' % P_ACTIVO].fill = PatternFill('solid', fgColor=GRIS)

    motor.val(ws, 'A13',
              'Estructura española de referencia: coste de producto en torno '
              'al 30 % y coste de personal del 30-35 % con servicio en mesa '
              '(15-25 % en barra o autoservicio). De ahí los dos objetivos: '
              '65 % y 55 %.')
    motor.val(ws, 'A14', 'Fuente: CaixaBankLab × elBulliFoundation, '
                         + URL_CAIXA)
    motor.val(ws, 'A15',
              'El 60 % que se cita a menudo como objetivo de prime cost es de '
              'Toast y está medido en Estados Unidos, con otra estructura de '
              'personal y otras cotizaciones sociales. Sirve de contraste, no '
              'de objetivo para un restaurante español.')
    motor.val(ws, 'A16',
              'Los dos objetivos son celdas verdes: si tu convenio, tu horario '
              'o tu modelo de servicio son distintos, cámbialos y el semáforo '
              'de la hoja «Mensual» se recalcula.')

    seccion(ws, 'A18', 'OTROS PARÁMETROS')
    cabecera(ws, 19, [('A', 'Parámetro'), ('B', 'Valor')], altura=20)
    motor.escribir_parametro(ws, P_SS, 'A', 'B', 'ss_empresa',
                             valor=D.SS_EMPRESA)
    motor.val(ws, 'A%d' % P_FC, 'Food cost objetivo (%)', bold=True)
    motor.val(ws, 'B%d' % P_FC, D.RESTAURANTE['food_cost_objetivo'],
              fmt=FMT_PCT, verde_=True)
    motor.val(ws, 'A23', motor.PARAMETROS['ss_empresa']['nota'])
    motor.val(ws, 'A24',
              'El food cost objetivo es el de tu casa. La media del sector en '
              'España ronda el 30 % (misma fuente), con la comida en el 28 % y '
              'la bebida en el 34,5 % sobre sus respectivos ingresos.')

    motor.dv_lista(ws, ['B%d' % P_TIPO], TIPOS_NEGOCIO,
                   titulo='Tipo de negocio')
    motor.dv_porcentaje(ws, ['B%d' % r for r in range(P_INI, P_FIN + 1)]
                        + ['B%d' % P_SS, 'B%d' % P_FC],
                        titulo='Porcentaje',
                        prompt='Se escribe en tanto por uno: 0,65 = 65 %.')
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
COLS_MENSUAL = [
    ('A', 'Mes'),
    ('B', 'Ventas netas de comida (€)'),
    ('C', 'Ventas netas de bebida (€)'),
    ('D', 'Ventas netas totales (€)'),
    ('E', 'Stock inicial de materia prima (€)'),
    ('F', 'Compras del mes (€)'),
    ('G', 'Stock final de materia prima (€)'),
    ('H', 'Consumo de materia prima (€)'),
    ('I', 'Coste de materia prima sobre ventas totales (%)'),
    ('J', 'Salarios brutos (€)'),
    ('K', 'Otros costes de personal (€)'),
    ('L', 'Coste de personal con Seguridad Social (€)'),
    ('M', 'Labor cost (%)'),
    ('N', 'Prime cost (%)'),
    ('O', 'Objetivo de prime cost (%)'),
    ('P', 'Margen tras prime cost (€)'),
    ('Q', 'Margen tras prime cost (%)'),
    ('R', 'Lectura del prime cost'),
]


def hoja_mensual(wb):
    ws = wb.create_sheet('Mensual')
    encabezar(ws, 'Prime cost mes a mes',
              nota='Todas las cifras van sin IVA. El consumo es stock inicial '
                   '+ compras − stock final, no las compras del mes.')
    cabecera(ws, M_CAB, COLS_MENSUAL, altura=62)
    anchos(ws, dict(zip('ABCDEFGHIJKLMNOPQR',
                        (12, 16, 16, 16, 16, 15, 16, 17, 12, 15, 16, 18, 12,
                         12, 14, 17, 16, 24))))

    ss = 'Parámetros!$B$%d' % P_SS
    for i, mes in enumerate(D.MESES):
        r = M_INI + i
        (v_com, v_beb, ini, compras, fin, brutos, otros) = D.CUADRO_MENSUAL[i]
        motor.val(ws, 'A%d' % r, mes, verde_=True)
        motor.val(ws, 'B%d' % r, float(v_com), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'C%d' % r, float(v_beb), fmt=FMT_EUR, verde_=True)
        motor.f(ws, 'D%d' % r,
                '=IFERROR(IF(AND($B{r}="",$C{r}=""),"",IF($B{r}="",0,$B{r})'
                '+IF($C{r}="",0,$C{r})),"")'.format(r=r), fmt=FMT_EUR)
        motor.val(ws, 'E%d' % r, float(ini), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'F%d' % r, float(compras), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'G%d' % r, float(fin), fmt=FMT_EUR, verde_=True)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF(OR($E{r}="",$F{r}="",$G{r}=""),"",'
                '$E{r}+$F{r}-$G{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF(OR($H{r}="",$D{r}="",$D{r}=0),"",$H{r}/$D{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.val(ws, 'J%d' % r, float(brutos), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'K%d' % r, float(otros), fmt=FMT_EUR, verde_=True)
        motor.f(ws, 'L%d' % r,
                '=IFERROR(IF($J{r}="","",$J{r}*(1+{ss})+IF($K{r}="",0,$K{r})),'
                '"")'.format(r=r, ss=ss), fmt=FMT_EUR)
        motor.f(ws, 'M%d' % r,
                '=IFERROR(IF(OR($L{r}="",$D{r}="",$D{r}=0),"",$L{r}/$D{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'N%d' % r,
                '=IFERROR(IF(OR($H{r}="",$L{r}="",$D{r}="",$D{r}=0),"",'
                '($H{r}+$L{r})/$D{r}),"")'.format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'O%d' % r,
                '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'
                .format(a=P_ACTIVO), fmt=FMT_PCT)
        motor.f(ws, 'P%d' % r,
                '=IFERROR(IF(OR($D{r}="",$H{r}="",$L{r}=""),"",'
                '$D{r}-$H{r}-$L{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'Q%d' % r,
                '=IFERROR(IF(OR($P{r}="",$D{r}="",$D{r}=0),"",$P{r}/$D{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'R%d' % r,
                '=IF(OR($N{r}="",$O{r}=""),"",IF($N{r}<=$O{r},"En objetivo",'
                '"Por encima del objetivo"))'.format(r=r))

    motor.val(ws, 'A%d' % M_TOT, 'TOTAL / MEDIA', bold=True)
    for col in ('B', 'C', 'D', 'F', 'H', 'J', 'K', 'L', 'P'):
        motor.f(ws, '%s%d' % (col, M_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=M_INI, b=M_FIN), fmt=FMT_EUR, bold=True)
    for col, num in (('I', 'H'), ('M', 'L'), ('Q', 'P')):
        motor.f(ws, '%s%d' % (col, M_TOT),
                '=IFERROR(IF(OR(${n}${t}="",$D${t}="",$D${t}=0),"",'
                '${n}${t}/$D${t}),"")'.format(n=num, t=M_TOT), fmt=FMT_PCT,
                bold=True)
    motor.f(ws, 'N%d' % M_TOT,
            '=IFERROR(IF(OR($H${t}="",$L${t}="",$D${t}="",$D${t}=0),"",'
            '($H${t}+$L${t})/$D${t}),"")'.format(t=M_TOT), fmt=FMT_PCT,
            bold=True)
    motor.f(ws, 'O%d' % M_TOT,
            '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'.format(a=P_ACTIVO),
            fmt=FMT_PCT, bold=True)
    motor.f(ws, 'R%d' % M_TOT,
            '=IF(OR($N${t}="",$O${t}=""),"",IF($N${t}<=$O${t},"En objetivo",'
            '"Por encima del objetivo"))'.format(t=M_TOT), bold=True)
    fila_total(ws, M_TOT, 'A', 'R')
    motor.val(ws, 'A%d' % (M_TOT + 1),
              'En la fila TOTAL no se suman los stocks: el consumo del año es '
              'la suma de los consumos de los doce meses, y los porcentajes '
              'son medias ponderadas sobre las ventas del año.')

    # --- espejo de parámetros (el formato condicional no cruza de hoja) ---
    seccion(ws, 'A%d' % M_ESP, 'PARÁMETROS EN VIGOR — se leen de la hoja '
                               '«Parámetros»; cámbialos allí, no aquí')
    motor.val(ws, 'A%d' % (M_ESP + 1), 'Tipo de negocio')
    motor.f(ws, 'E%d' % (M_ESP + 1), '=Parámetros!$B$%d' % P_TIPO)
    motor.val(ws, 'A%d' % (M_ESP + 2), 'Objetivo de prime cost (%)')
    motor.f(ws, 'E%d' % (M_ESP + 2),
            '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'.format(a=P_ACTIVO),
            fmt=FMT_PCT)
    motor.val(ws, 'A%d' % (M_ESP + 3), 'Food cost objetivo (%)')
    motor.f(ws, 'E%d' % (M_ESP + 3),
            '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'.format(a=P_FC),
            fmt=FMT_PCT)
    motor.val(ws, 'A%d' % (M_ESP + 4),
              'Seguridad Social a cargo de la empresa (%)')
    motor.f(ws, 'E%d' % (M_ESP + 4),
            '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'.format(a=P_SS),
            fmt=FMT_PCT)
    for i in range(1, 5):
        ws['E%d' % (M_ESP + i)].fill = PatternFill('solid', fgColor=GRIS)

    # --- semáforos ---
    fc_obj = '$E$%d' % (M_ESP + 3)
    cf_expresion(ws, 'N{a}:N{b}'.format(a=M_INI, b=M_TOT),
                 '=AND(ISNUMBER($N{a}),ISNUMBER($O{a}),$N{a}>$O{a})'
                 .format(a=M_INI), motor.CF_ROJO_BG, motor.CF_ROJO_FG)
    cf_expresion(ws, 'N{a}:N{b}'.format(a=M_INI, b=M_TOT),
                 '=AND(ISNUMBER($N{a}),ISNUMBER($O{a}),$N{a}<=$O{a})'
                 .format(a=M_INI), motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    cf_expresion(ws, 'I{a}:I{b}'.format(a=M_INI, b=M_TOT),
                 '=AND(ISNUMBER($I{a}),ISNUMBER({o}),$I{a}>{o})'
                 .format(a=M_INI, o=fc_obj),
                 motor.CF_ROJO_BG, motor.CF_ROJO_FG)
    cf_expresion(ws, 'I{a}:I{b}'.format(a=M_INI, b=M_TOT),
                 '=AND(ISNUMBER($I{a}),ISNUMBER({o}),$I{a}<={o})'
                 .format(a=M_INI, o=fc_obj),
                 motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    motor.semaforo_texto(ws, 'R{a}:R{b}'.format(a=M_INI, b=M_TOT),
                         (('En objetivo', motor.CF_VERDE_BG,
                           motor.CF_VERDE_FG),
                          ('Por encima del objetivo', motor.CF_ROJO_BG,
                           motor.CF_ROJO_FG)))

    # --- validaciones ---
    filas = list(range(M_INI, M_FIN + 1))
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'BCEFGJK' for r in filas],
                      minimo=0, titulo='Importe (€)')

    # --- gráfico de líneas: prime cost vs objetivo ---
    grafico = LineChart()
    grafico.title = 'Prime cost mensual frente al objetivo'
    grafico.style = 2
    grafico.height = 8.5
    grafico.width = 24
    datos = Reference(ws, min_col=14, max_col=15, min_row=M_CAB, max_row=M_FIN)
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(Reference(ws, min_col=1, min_row=M_INI,
                                     max_row=M_FIN))
    grafico.y_axis.numFmt = FMT_PCT
    grafico.y_axis.title = 'Prime cost (%)'
    grafico.x_axis.title = 'Mes'
    grafico.x_axis.delete = False
    grafico.y_axis.delete = False
    grafico.x_axis.tickLblPos = 'low'
    grafico.y_axis.majorGridlines = None
    ws.add_chart(grafico, 'G%d' % M_ESP)

    ws.freeze_panes = 'B5'
    pagina(ws, titulos='$%d:$%d' % (M_CAB, M_CAB))
    return ws


# --------------------------------------------------------------------------
def mapa():
    celdas_p = {
        'Tipo de negocio elegido': 'B%d' % P_TIPO,
        'Objetivo de prime cost con servicio en mesa': 'B%d' % P_INI,
        'Objetivo de prime cost en barra o autoservicio': 'B%d' % P_FIN,
        'Objetivo de prime cost en vigor': 'B%d' % P_ACTIVO,
        'Seguridad Social a cargo de la empresa': 'B%d' % P_SS,
        'Food cost objetivo': 'B%d' % P_FC,
    }
    celdas_m = {
        'Ventas netas de comida del año': 'B%d' % M_TOT,
        'Ventas netas de bebida del año': 'C%d' % M_TOT,
        'Ventas netas totales del año': 'D%d' % M_TOT,
        'Compras del año': 'F%d' % M_TOT,
        'Consumo de materia prima del año': 'H%d' % M_TOT,
        'Coste de materia prima sobre ventas totales del año (ponderado)':
            'I%d' % M_TOT,
        'Salarios brutos del año': 'J%d' % M_TOT,
        'Otros costes de personal del año': 'K%d' % M_TOT,
        'Coste de personal con Seguridad Social del año': 'L%d' % M_TOT,
        'Labor cost del año (ponderado)': 'M%d' % M_TOT,
        'Prime cost del año (ponderado)': 'N%d' % M_TOT,
        'Objetivo de prime cost en vigor (espejo)': 'O%d' % M_TOT,
        'Margen tras prime cost del año': 'P%d' % M_TOT,
        'Margen tras prime cost del año (%)': 'Q%d' % M_TOT,
        'Lectura del prime cost del año': 'R%d' % M_TOT,
    }
    for i, mes in enumerate(D.MESES):
        r = M_INI + i
        celdas_m['Ventas netas totales de %s' % mes] = 'D%d' % r
        celdas_m['Consumo de materia prima de %s' % mes] = 'H%d' % r
        celdas_m['Coste de materia prima sobre ventas totales de %s' % mes] = \
            'I%d' % r
        celdas_m['Coste de personal con SS de %s' % mes] = 'L%d' % r
        celdas_m['Labor cost de %s' % mes] = 'M%d' % r
        celdas_m['Prime cost de %s' % mes] = 'N%d' % r
        celdas_m['Margen tras prime cost de %s' % mes] = 'P%d' % r
        celdas_m['Lectura del prime cost de %s' % mes] = 'R%d' % r
    return {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Parámetros': {
                'celdas': celdas_p,
                'tablas': [{'titulo': 'Objetivo de prime cost por tipo de '
                                      'negocio',
                            'cols': [['Tipo de negocio', 'A', 'txt'],
                                     ['Objetivo de prime cost (%)',
                                      'B', 'pct1']],
                            'filas': [P_INI, P_FIN]}],
            },
            'Mensual': {
                'celdas': celdas_m,
                'tablas': [{'titulo': 'Cuadro de mando mensual del prime cost',
                            'cols': [[t, c, f] for (c, t), f in zip(
                                COLS_MENSUAL,
                                ['txt', 'eur', 'eur', 'eur', 'eur', 'eur',
                                 'eur', 'eur', 'pct1', 'eur', 'eur', 'eur',
                                 'pct1', 'pct1', 'pct1', 'eur', 'pct1',
                                 'txt'])],
                            'filas': [M_INI, M_TOT]}],
            },
        },
    }


def main():
    import json
    wb = Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_parametros(wb)
    hoja_mensual(wb)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · Versión 1.0 · septiembre 2026'

    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    destino = os.path.join(AQUI, 'build')
    os.makedirs(destino, exist_ok=True)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')
    wb.save(ruta)
    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)
    print('escrito:', ruta)
    print('fórmulas registradas:', len(motor.REGISTRO))


if __name__ == '__main__':
    main()
