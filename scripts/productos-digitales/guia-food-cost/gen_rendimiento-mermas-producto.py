#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_rendimiento-mermas-producto.py — libro 2 de los 8 del producto «Guía
Food Cost + Ingeniería de Menú» (SPEC §2.2, fila 2).

Tres hojas de trabajo: el test de rendimiento con subproductos aprovechables,
la merma de cocción y «Mi Tabla de Mermas», que sustituye la tabla genérica por
la medición propia en cuanto existe.

Cifras: TODAS de `datos_ejemplo` (TESTS_RENDIMIENTO, TESTS_COCCION,
MERMAS_REFERENCIA, FUENTE_MERMAS). Las mermas «medidas» de la tercera hoja se
DERIVAN de los tests de la primera; no se teclea ningún número nuevo.

Sale SIEMPRE en `build/rendimiento-mermas-producto.xlsx`. Después:
    python3 ../inject_cache.py build/rendimiento-mermas-producto.xlsx
"""
import io
import json
import os
import sys

AQUI = os.path.dirname(os.path.abspath(__file__))
GUIAS = '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0'
sys.path.insert(0, GUIAS)
sys.path.insert(0, AQUI)

from openpyxl import Workbook                                    # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill         # noqa: E402
from openpyxl.utils import (column_index_from_string,            # noqa: E402
                            get_column_letter)
from openpyxl.worksheet.page import PageMargins                  # noqa: E402
from openpyxl.worksheet.properties import PageSetupProperties    # noqa: E402

import motor                                                     # noqa: E402
import datos_ejemplo as D                                        # noqa: E402

PID = 'guia-food-cost-ingenieria-menu'
PRODUCTO = 'Guía Food Cost + Ingeniería de Menú'
NOMBRE = 'rendimiento-mermas-producto'
TITULO = 'Rendimiento, Mermas y Coste Neto del Producto'
SUBTITULO = 'AI Chef Pro · aichef.pro — Guía Food Cost + Ingeniería de Menú'
VERSION = ('Versión 1.0 · septiembre 2026 · aichef.pro/'
           + PID + ' · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 '
       '· johnguerrero.es')
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde, '
               'desprotege la hoja (sin contraseña).')
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

motor.CTX['producto'] = PID

CAB = '2D2D2D'
CREMA = 'FFF8DC'
FMT_EUR = motor.FMT_EUR
FMT_PCT = motor.FMT_PCT
FMT_CANT = '0.000'
FMT_ENT = '#,##0'

TIPO_POR_FMT = {FMT_EUR: 'eur', FMT_PCT: 'pct1', FMT_CANT: 'num1',
                FMT_ENT: 'num', None: 'txt'}

# --------------------------------------------------------------------------
# Hoja 1 · Test de Rendimiento
# --------------------------------------------------------------------------
R_CAB, R_INI, R_N = 5, 6, 15
R_FIN = R_INI + R_N - 1
COLS_R = [
    ('A', '#', 5, FMT_ENT, False),
    ('B', 'Producto', 32, None, True),
    ('C', 'Peso bruto (kg)', 12, FMT_CANT, True),
    ('D', 'Precio/kg bruto sin IVA (€)', 14, FMT_EUR, True),
    ('E', 'Peso limpio (kg)', 12, FMT_CANT, True),
    ('F', 'Subproductos aprovechables (kg)', 14, FMT_CANT, True),
    ('G', 'Valor de uso del subproducto (€/kg)', 14, FMT_EUR, True),
    ('H', 'Rendimiento (%)', 12, FMT_PCT, False),
    ('I', 'Merma (%)', 11, FMT_PCT, False),
    ('J', 'Factor de corrección (1/rendimiento)', 13, FMT_CANT, False),
    ('K', 'Coste de compra (€)', 13, FMT_EUR, False),
    ('L', 'Valor de los subproductos (€)', 13, FMT_EUR, False),
    ('M', 'Coste neto €/kg limpio SIN aprovechar', 15, FMT_EUR, False),
    ('N', 'Coste neto €/kg limpio APROVECHANDO', 15, FMT_EUR, False),
    ('O', 'Ahorro por aprovechar (€/kg)', 14, FMT_EUR, False),
    ('P', 'Sobrecoste sobre el precio bruto (€/kg)', 14, FMT_EUR, False),
]

# --------------------------------------------------------------------------
# Hoja 2 · Merma de Cocción
# --------------------------------------------------------------------------
C_CAB, C_INI, C_N = 5, 6, 10
C_FIN = C_INI + C_N - 1
COLS_C = [
    ('A', '#', 5, FMT_ENT, False),
    ('B', 'Elaboración', 34, None, True),
    ('C', 'Técnica', 14, None, True),
    ('D', 'Peso crudo (kg)', 13, FMT_CANT, True),
    ('E', 'Peso cocinado (kg)', 13, FMT_CANT, True),
    ('F', 'Rendimiento de cocción (%)', 13, FMT_PCT, False),
    ('G', 'Pérdida por cocción (%)', 13, FMT_PCT, False),
    ('H', 'Factor de cocción (1/rendimiento)', 13, FMT_CANT, False),
    ('I', 'Coste/kg del producto crudo (€)', 14, FMT_EUR, True),
    ('J', 'Coste/kg del producto ya cocinado (€)', 15, FMT_EUR, False),
    ('K', 'Sobrecoste por la cocción (€/kg)', 14, FMT_EUR, False),
    ('L', 'Notas', 30, None, True),
]
TECNICAS = ['Plancha', 'Horno', 'Brasa', 'Confitado', 'Hervido', 'Vapor',
            'Fritura', 'Guisado', 'Baja temperatura']

# --------------------------------------------------------------------------
# Hoja 3 · Mi Tabla de Mermas
# --------------------------------------------------------------------------
M_CAB, M_INI, M_N = 5, 6, 18
M_FIN = M_INI + M_N - 1
COLS_M = [
    ('A', '#', 5, FMT_ENT, False),
    ('B', 'Categoría de producto', 34, None, True),
    ('C', 'Merma de referencia mínima (%)', 14, FMT_PCT, True),
    ('D', 'Merma de referencia máxima (%)', 14, FMT_PCT, True),
    ('E', 'Merma de referencia media (%)', 14, FMT_PCT, False),
    ('F', 'Tu merma medida (%)', 13, FMT_PCT, True),
    ('G', 'Merma que usas (%)', 13, FMT_PCT, False),
    ('H', 'De dónde sale el dato', 20, None, False),
    ('I', 'Diferencia con la referencia (puntos)', 15, FMT_PCT, False),
    ('J', 'Nota', 40, None, True),
]

#: Categoría de la tabla de mermas → índice del test de rendimiento que la
#: mide. Así la columna «Tu merma medida» sale de los tests de la hoja 1 y no
#: de un número inventado.
MEDIDA_DESDE_TEST = {
    'Solomillo de vacuno (limpieza)': 2,
    'Pescado entero': 0,
    'Alcachofa': 6,
    'Mejillones, caracoles, callos': 8,
    'Marisco entero (gamba, cigala)': 9,
    'Aves enteras': 3,
}


# --------------------------------------------------------------------------
# Utilidades de presentación
# --------------------------------------------------------------------------
def a4(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.51, right=0.51, top=0.55, bottom=0.55,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def cabecera_libro(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color='595959')


def fila_cabecera(ws, fila, columnas):
    for letra, texto, _, _, _ in columnas:
        cel = motor.val(ws, letra + str(fila), texto, bold=True, wrap=True)
        cel.fill = PatternFill('solid', fgColor=CAB)
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    ws.row_dimensions[fila].height = 42


def _letras(desde, hasta):
    return [get_column_letter(i)
            for i in range(column_index_from_string(desde),
                           column_index_from_string(hasta) + 1)]


def banda(ws, fila, desde, hasta, texto):
    motor.val(ws, desde + str(fila), texto, bold=True)
    for col in _letras(desde, hasta):
        ws[col + str(fila)].fill = PatternFill('solid', fgColor=CREMA)
        ws[col + str(fila)].font = Font(bold=True, size=11)


def preparar(ws, columnas, fila_cab, fila_ini, fila_fin, titulo):
    cabecera_libro(ws, titulo)
    motor.anchos(ws, dict((c[0], c[2]) for c in columnas))
    fila_cabecera(ws, fila_cab, columnas)
    ws.freeze_panes = 'C' + str(fila_ini)
    ws.auto_filter.ref = ('A' + str(fila_cab) + ':' + columnas[-1][0]
                          + str(fila_fin))
    for i in range(fila_ini, fila_fin + 1):
        motor.val(ws, 'A' + str(i), i - fila_ini + 1, fmt=FMT_ENT,
                  align='center')
        for letra, _, _, fmt, entrada in columnas:
            if letra == 'A':
                continue
            if fmt:
                ws[letra + str(i)].number_format = fmt
            if entrada:
                motor.verde(ws, letra + str(i))


# --------------------------------------------------------------------------
def hoja_rendimiento(wb):
    ws = wb.create_sheet('Test de Rendimiento')
    preparar(ws, COLS_R, R_CAB, R_INI, R_FIN, TITULO)
    motor.val(ws, 'B3', 'Pesa el producto entero, límpialo, pesa lo que queda '
                        'y pesa aparte lo que sí vas a usar (espinas, cabezas, '
                        'carcasas, recortes).', wrap=True)

    for i in range(R_N):
        r = R_INI + i
        if i < len(D.TESTS_RENDIMIENTO):
            nom, bruto, precio, limpio, sub_kg, sub_val = D.TESTS_RENDIMIENTO[i]
            motor.val(ws, 'B' + str(r), nom)
            motor.val(ws, 'C' + str(r), bruto, fmt=FMT_CANT)
            motor.val(ws, 'D' + str(r), precio, fmt=FMT_EUR)
            motor.val(ws, 'E' + str(r), limpio, fmt=FMT_CANT)
            motor.val(ws, 'F' + str(r), sub_kg, fmt=FMT_CANT)
            motor.val(ws, 'G' + str(r), sub_val, fmt=FMT_EUR)
        motor.f(ws, 'H' + str(r),
                '=IFERROR(IF(OR($C{r}="",$E{r}="",$C{r}=0),"",$E{r}/$C{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'I' + str(r),
                '=IFERROR(IF($H{r}="","",1-$H{r}),"")'.format(r=r),
                fmt=FMT_PCT)
        motor.f(ws, 'J' + str(r),
                '=IFERROR(IF(OR($H{r}="",$H{r}=0),"",1/$H{r}),"")'.format(r=r),
                fmt=FMT_CANT)
        motor.f(ws, 'K' + str(r),
                '=IFERROR(IF(OR($C{r}="",$D{r}=""),"",$C{r}*$D{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'L' + str(r),
                '=IFERROR(IF(OR($F{r}="",$G{r}=""),"",$F{r}*$G{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'M' + str(r),
                '=IFERROR(IF(OR($K{r}="",$E{r}="",$E{r}=0),"",$K{r}/$E{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'N' + str(r),
                '=IFERROR(IF(OR($K{r}="",$E{r}="",$E{r}=0),"",'
                '($K{r}-IF($L{r}="",0,$L{r}))/$E{r}),"")'.format(r=r),
                fmt=FMT_EUR)
        motor.f(ws, 'O' + str(r),
                '=IFERROR(IF(OR($M{r}="",$N{r}=""),"",$M{r}-$N{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'P' + str(r),
                '=IFERROR(IF(OR($N{r}="",$D{r}=""),"",$N{r}-$D{r}),"")'
                .format(r=r), fmt=FMT_EUR)

    r = R_FIN + 2
    banda(ws, r, 'B', 'H', 'RESUMEN — lo calcula el libro')
    celdas = {}

    r += 1
    celdas['Productos medidos'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Productos medidos (n.º)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($E${a}:$E${b},">0")=0,"",COUNTIF($E${a}:$E${b},">0"))'
            .format(a=R_INI, b=R_FIN), fmt=FMT_ENT)

    r += 1
    celdas['Peso bruto total comprado'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Peso bruto total comprado (kg)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($C${a}:$C${b},">0")=0,"",SUM($C${a}:$C${b}))'
            .format(a=R_INI, b=R_FIN), fmt=FMT_CANT)

    r += 1
    f_limpio = r
    celdas['Peso limpio total'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Peso limpio total (kg)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($E${a}:$E${b},">0")=0,"",SUM($E${a}:$E${b}))'
            .format(a=R_INI, b=R_FIN), fmt=FMT_CANT)

    r += 1
    f_rend = r
    celdas['Rendimiento medio ponderado'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Rendimiento medio ponderado (%)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(SUM($C${a}:$C${b})=0,"",SUM($E${a}:$E${b})'
            '/SUM($C${a}:$C${b})),"")'.format(a=R_INI, b=R_FIN),
            fmt=FMT_PCT, bold=True)

    r += 1
    celdas['Merma media ponderada'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Merma media ponderada (%)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF($E${x}="","",1-$E${x}),"")'.format(x=f_rend),
            fmt=FMT_PCT, bold=True)

    r += 1
    celdas['Coste total de compra'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Coste total de compra (€)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($K${a}:$K${b},">0")=0,"",SUM($K${a}:$K${b}))'
            .format(a=R_INI, b=R_FIN), fmt=FMT_EUR)

    r += 1
    f_sub = r
    celdas['Valor de uso de los subproductos'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Valor de uso de los subproductos aprovechables (€)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($L${a}:$L${b},">0")=0,"",SUM($L${a}:$L${b}))'
            .format(a=R_INI, b=R_FIN), fmt=FMT_EUR)

    r += 1
    celdas['Ahorro medio por aprovechar'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Ahorro medio por aprovechar los subproductos (€/kg limpio)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(OR($E${s}="",$E${l}="",$E${l}=0),"",$E${s}/$E${l}),"")'
            .format(s=f_sub, l=f_limpio), fmt=FMT_EUR)
    motor.val(ws, 'I' + str(r),
              'Si este ahorro no compensa las horas de despiece, compra el '
              'producto ya limpio y quédate con el rendimiento del proveedor.',
              wrap=True)

    filas = [str(x) for x in range(R_INI, R_FIN + 1)]
    for letra, titulo, prompt in (
            ('C', 'Peso bruto', 'El peso del producto tal y como llega, en kg.'),
            ('E', 'Peso limpio', 'Lo que queda listo para porcionar, en kg.'),
            ('F', 'Subproductos', 'Kilos aprovechables: espinas, cabezas, '
                                  'carcasas, recortes.')):
        motor.dv_numerica(ws, [letra + x for x in filas], minimo=0,
                          titulo=titulo, prompt=prompt)
    for letra, titulo, prompt in (
            ('D', 'Precio de compra', 'Precio por kg BRUTO sin IVA, el del '
                                      'albarán.'),
            ('G', 'Valor del subproducto', 'Lo que te costaría comprar ese '
                                           'subproducto hecho (un fumet, un '
                                           'caldo): es su valor de uso.')):
        motor.dv_numerica(ws, [letra + x for x in filas], minimo=0,
                          titulo=titulo, prompt=prompt)

    a4(ws, apaisado=True, titulos=str(R_CAB) + ':' + str(R_CAB))
    return celdas


# --------------------------------------------------------------------------
def hoja_coccion(wb):
    ws = wb.create_sheet('Merma de Cocción')
    preparar(ws, COLS_C, C_CAB, C_INI, C_FIN, 'Merma de Cocción')
    motor.val(ws, 'B3', 'Pesa la pieza cruda y vuelve a pesarla al salir de la '
                        'cocción, ya reposada. La pérdida de cocción no se ve '
                        'en el albarán y sí en el plato.', wrap=True)

    # El coste/kg crudo de las 5 pruebas, siempre en NETO €/kg limpio (la
    # nota de M6/M7 de esta misma hoja lo pide: si el producto pasa antes por
    # despiece, no vale el precio bruto de albarán). El solomillo iba mal:
    # 15,80 €/kg es el precio BRUTO de la ficha (línea 1, sin ajustar por
    # merma); el neto es bruto / (1 - merma) = 15,80 / (1 - 0,12) = 17,95
    # €/kg, el mismo criterio que aplica la propia ficha de escandallo.
    _solomillo_bruto, _solomillo_merma = (D.FICHA['lineas'][0][3],
                                          D.FICHA['lineas'][0][4])
    coste_crudo = {
        'Solomillo de cerdo a la plancha':
            round(_solomillo_bruto / (1 - _solomillo_merma), 2),
        'Pollo de corral al horno': D.TESTS_RENDIMIENTO[3][2],
        # Bacalao desalado de calidad similar a la lubina entera de la hoja
        # «Test de Rendimiento» (14,90 €/kg); secreto ibérico, corte premium
        # de cerdo (16,50 €/kg); verduras de temporada para asar, precio
        # medio de mercado (2,20 €/kg).
        'Bacalao confitado': 14.90,
        'Secreto ibérico a la brasa': 16.50,
        'Verduras asadas': 2.20,
    }
    for i in range(C_N):
        r = C_INI + i
        if i < len(D.TESTS_COCCION):
            nom, tec, crudo, cocinado = D.TESTS_COCCION[i]
            motor.val(ws, 'B' + str(r), nom)
            motor.val(ws, 'C' + str(r), tec, align='center')
            motor.val(ws, 'D' + str(r), crudo, fmt=FMT_CANT)
            motor.val(ws, 'E' + str(r), cocinado, fmt=FMT_CANT)
            if nom in coste_crudo:
                motor.val(ws, 'I' + str(r), coste_crudo[nom], fmt=FMT_EUR)
        motor.f(ws, 'F' + str(r),
                '=IFERROR(IF(OR($D{r}="",$E{r}="",$D{r}=0),"",$E{r}/$D{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'G' + str(r),
                '=IFERROR(IF($F{r}="","",1-$F{r}),"")'.format(r=r),
                fmt=FMT_PCT)
        motor.f(ws, 'H' + str(r),
                '=IFERROR(IF(OR($F{r}="",$F{r}=0),"",1/$F{r}),"")'.format(r=r),
                fmt=FMT_CANT)
        motor.f(ws, 'J' + str(r),
                '=IFERROR(IF(OR($I{r}="",$F{r}="",$F{r}=0),"",$I{r}/$F{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'K' + str(r),
                '=IFERROR(IF(OR($J{r}="",$I{r}=""),"",$J{r}-$I{r}),"")'
                .format(r=r), fmt=FMT_EUR)

    motor.val(ws, 'M' + str(C_INI),
              'Escribe en la columna I el coste por kg del producto crudo. Si '
              'el producto pasa antes por despiece, usa el «Coste neto €/kg '
              'limpio APROVECHANDO» de la hoja «Test de Rendimiento».',
              wrap=True)

    r = C_FIN + 2
    banda(ws, r, 'B', 'H', 'RESUMEN — lo calcula el libro')
    celdas = {}

    r += 1
    f_rend = r
    celdas['Rendimiento medio de coccion'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Rendimiento medio de cocción ponderado (%)',
              bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(SUM($D${a}:$D${b})=0,"",SUM($E${a}:$E${b})'
            '/SUM($D${a}:$D${b})),"")'.format(a=C_INI, b=C_FIN),
            fmt=FMT_PCT, bold=True)

    r += 1
    celdas['Perdida media de coccion'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Pérdida media por cocción (%)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF($E${x}="","",1-$E${x}),"")'.format(x=f_rend),
            fmt=FMT_PCT, bold=True)

    r += 1
    celdas['Pruebas de coccion registradas'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Pruebas de cocción registradas (n.º)')
    motor.f(ws, 'E' + str(r),
            '=IF(COUNTIF($E${a}:$E${b},">0")=0,"",COUNTIF($E${a}:$E${b},">0"))'
            .format(a=C_INI, b=C_FIN), fmt=FMT_ENT)

    filas = [str(x) for x in range(C_INI, C_FIN + 1)]
    motor.dv_lista(ws, ['C' + x for x in filas], TECNICAS,
                   titulo='Técnica no válida')
    motor.dv_numerica(ws, ['D' + x for x in filas], minimo=0,
                      titulo='Peso crudo', prompt='Peso antes de cocinar, en '
                                                  'kg.')
    motor.dv_numerica(ws, ['E' + x for x in filas], minimo=0,
                      titulo='Peso cocinado',
                      prompt='Peso después de cocinar y reposar, en kg.')
    motor.dv_numerica(ws, ['I' + x for x in filas], minimo=0,
                      titulo='Coste del producto crudo',
                      prompt='Coste por kg del producto crudo, sin IVA.')

    a4(ws, apaisado=True, titulos=str(C_CAB) + ':' + str(C_CAB))
    return celdas


# --------------------------------------------------------------------------
def hoja_mermas(wb):
    ws = wb.create_sheet('Mi Tabla de Mermas')
    preparar(ws, COLS_M, M_CAB, M_INI, M_FIN, 'Mi Tabla de Mermas')
    motor.val(ws, 'B3', 'La referencia solo es el punto de partida. En cuanto '
                        'escribes tu merma medida, el libro deja de usar la '
                        'referencia y usa la tuya.', wrap=True)

    for i in range(M_N):
        r = M_INI + i
        if i < len(D.MERMAS_REFERENCIA):
            cat, mn, mx, nota = D.MERMAS_REFERENCIA[i]
            motor.val(ws, 'B' + str(r), cat)
            motor.val(ws, 'C' + str(r), mn, fmt=FMT_PCT)
            motor.val(ws, 'D' + str(r), mx, fmt=FMT_PCT)
            if nota:
                motor.val(ws, 'J' + str(r), nota, wrap=True)
            idx = MEDIDA_DESDE_TEST.get(cat)
            if idx is not None:
                bruto = D.TESTS_RENDIMIENTO[idx][1]
                limpio = D.TESTS_RENDIMIENTO[idx][3]
                motor.val(ws, 'F' + str(r), round(1 - limpio / bruto, 4),
                          fmt=FMT_PCT)
        motor.f(ws, 'E' + str(r),
                '=IFERROR(IF(OR($C{r}="",$D{r}=""),"",($C{r}+$D{r})/2),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'G' + str(r),
                '=IF(ISNUMBER($F{r}),$F{r},IF($E{r}="","",$E{r}))'.format(r=r),
                fmt=FMT_PCT, bold=True)
        motor.f(ws, 'H' + str(r),
                '=IF(ISNUMBER($F{r}),"Tu medición",'
                'IF($E{r}="","","Referencia orientativa"))'.format(r=r))
        motor.f(ws, 'I' + str(r),
                '=IF(ISNUMBER($F{r}),IF($E{r}="","",$F{r}-$E{r}),"")'
                .format(r=r), fmt=FMT_PCT)

    motor.semaforo_texto(
        ws, 'H' + str(M_INI) + ':H' + str(M_FIN),
        (('Tu medición', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
         ('Referencia orientativa', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG)))

    r = M_FIN + 2
    banda(ws, r, 'B', 'H', 'RESUMEN — lo calcula el libro')
    celdas = {}

    r += 1
    celdas['Categorias con medicion propia'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Categorías con medición propia (n.º)',
              bold=True)
    motor.f(ws, 'E' + str(r),
            '=COUNTIF($H${a}:$H${b},"Tu medición")'.format(a=M_INI, b=M_FIN),
            fmt=FMT_ENT, bold=True)

    r += 1
    celdas['Categorias con referencia'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Categorías que aún usan la referencia (n.º)')
    motor.f(ws, 'E' + str(r),
            '=COUNTIF($H${a}:$H${b},"Referencia orientativa")'
            .format(a=M_INI, b=M_FIN), fmt=FMT_ENT)

    r += 1
    celdas['Desviacion media de tus mediciones'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Desviación media de tus mediciones sobre la referencia (puntos)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(COUNTIF($H${a}:$H${b},"Tu medición")=0,"",'
            'SUM($I${a}:$I${b})/COUNTIF($H${a}:$H${b},"Tu medición")),"")'
            .format(a=M_INI, b=M_FIN), fmt=FMT_PCT)
    motor.val(ws, 'J' + str(r),
              'Si tu merma está sistemáticamente por encima de la referencia, '
              'el problema no es la tabla: es el despiece, el proveedor o el '
              'calibre que te sirven.', wrap=True)

    filas = [str(x) for x in range(M_INI, M_FIN + 1)]
    for letra, titulo in (('C', 'Merma de referencia mínima'),
                          ('D', 'Merma de referencia máxima'),
                          ('F', 'Tu merma medida')):
        motor.dv_porcentaje(ws, [letra + x for x in filas], titulo=titulo,
                            prompt=motor.PCT_MERMA[1],
                            maximo=motor.PCT_MERMA[2])

    a4(ws, apaisado=True, titulos=str(M_CAB) + ':' + str(M_CAB))
    return celdas


# --------------------------------------------------------------------------
PASOS = [
    '1. Hoja «Test de Rendimiento»: pesa el producto entero, límpialo, pesa lo '
    'que queda y pesa aparte los subproductos que sí vas a usar.',
    '2. El libro te da el rendimiento, la merma, el factor de corrección y el '
    'coste NETO por kg limpio, con y sin aprovechar los subproductos.',
    '3. Compara ese coste neto con el precio del albarán: la diferencia es lo '
    'que de verdad te cuesta el kilo que sale al pase.',
    '4. Hoja «Merma de Cocción»: pesa la pieza cruda y ya cocinada y reposada; '
    'sale el coste por kg del producto terminado.',
    '5. Hoja «Mi Tabla de Mermas»: escribe tu merma medida y la columna «Merma '
    'que usas» cambia sola de la referencia a la tuya.',
    '6. Lleva la merma de la columna «Merma que usas» a la columna «Merma (%)» '
    'de ficha-escandallo-base.xlsx: ahí es donde se convierte en dinero.',
]

NOTA_IVA_LIBRO = (
    'Todos los precios de este libro van SIN IVA, igual que en el resto del '
    'pack: el IVA soportado de las compras se deduce en el modelo 303, así que '
    'es tesorería y no coste.')


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 112
    motor.val(ws, 'A1', TITULO, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color='595959')
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), 'De dónde salen las mermas de referencia',
              bold=True)
    fila += 1
    coords = {'fuente_mermas': 'A' + str(fila)}
    motor.val(ws, 'A' + str(fila), D.FUENTE_MERMAS, wrap=True)
    fila += 2
    for texto in (motor.NOTA_VERDES, NOTA_IVA_LIBRO, DESPROTEGER, BIO,
                  VERSION):
        cel = motor.val(ws, 'A' + str(fila), texto, wrap=True)
        if texto is motor.NOTA_VERDES:
            cel.fill = PatternFill('solid', fgColor=motor.VERDE)
        if texto is VERSION:
            coords['version'] = 'A' + str(fila)
        fila += 1
    a4(ws, apaisado=False)
    return coords


# --------------------------------------------------------------------------

def _fila_test(nombre):
    for i, t in enumerate(D.TESTS_RENDIMIENTO):
        if t[0] == nombre:
            return R_INI + i
    raise KeyError(nombre)


def _fila_coccion(nombre):
    for i, t in enumerate(D.TESTS_COCCION):
        if t[0] == nombre:
            return C_INI + i
    raise KeyError(nombre)


def _fila_merma(categoria):
    for i, t in enumerate(D.MERMAS_REFERENCIA):
        if t[0] == categoria:
            return M_INI + i
    raise KeyError(categoria)


def cols_mapa(columnas):
    return [[c[1], c[0], TIPO_POR_FMT[c[3]]] for c in columnas]


def escribir_mapa(cr, cc, cm, ci):
    mapa = {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Instrucciones': {
                'celdas': {'Título del libro': 'A1',
                           'Fuente de las mermas de referencia':
                               ci['fuente_mermas'],
                           'Línea de versión': ci['version']},
                'tablas': [],
            },
            'Test de Rendimiento': {
                'celdas': {
                    'Productos medidos': cr['Productos medidos'],
                    'Peso bruto total comprado': cr['Peso bruto total comprado'],
                    'Peso limpio total': cr['Peso limpio total'],
                    'Rendimiento medio ponderado':
                        cr['Rendimiento medio ponderado'],
                    'Merma media ponderada': cr['Merma media ponderada'],
                    'Coste total de compra': cr['Coste total de compra'],
                    'Valor de uso de los subproductos':
                        cr['Valor de uso de los subproductos'],
                    'Ahorro medio por aprovechar los subproductos':
                        cr['Ahorro medio por aprovechar'],
                    # Anclas de fila que citan los capítulos 05 y 19.
                    'Rendimiento de la lubina entera':
                        'H' + str(_fila_test('Lubina entera (1,2 kg)')),
                    'Coste neto por kg limpio de la lubina, aprovechando':
                        'N' + str(_fila_test('Lubina entera (1,2 kg)')),
                    'Sobrecoste de la lubina sobre el precio bruto':
                        'P' + str(_fila_test('Lubina entera (1,2 kg)')),
                    'Rendimiento de la alcachofa':
                        'H' + str(_fila_test('Alcachofa')),
                    'Rendimiento del mejillón con concha':
                        'H' + str(_fila_test('Mejillón (con concha)')),
                    'Ahorro por aprovechar las cabezas de la gamba':
                        'O' + str(_fila_test('Gamba blanca (entera)')),
                },
                'tablas': [
                    {'titulo': 'Tests de rendimiento por producto',
                     'cols': cols_mapa(COLS_R),
                     'filas': [R_INI, R_FIN]},
                ],
            },
            'Merma de Cocción': {
                'celdas': {
                    'Rendimiento medio de cocción ponderado':
                        cc['Rendimiento medio de coccion'],
                    'Pérdida media por cocción':
                        cc['Perdida media de coccion'],
                    'Pruebas de cocción registradas':
                        cc['Pruebas de coccion registradas'],
                    'Pérdida de cocción del pollo de corral al horno':
                        'G' + str(_fila_coccion('Pollo de corral al horno')),
                    'Coste por kg del pollo ya cocinado':
                        'J' + str(_fila_coccion('Pollo de corral al horno')),
                    'Pérdida de cocción de las verduras asadas':
                        'G' + str(_fila_coccion('Verduras asadas')),
                    'Pérdida de cocción del solomillo de cerdo a la plancha':
                        'G' + str(_fila_coccion(
                            'Solomillo de cerdo a la plancha')),
                },
                'tablas': [
                    {'titulo': 'Pruebas de merma de cocción',
                     'cols': cols_mapa(COLS_C),
                     'filas': [C_INI, C_FIN]},
                ],
            },
            'Mi Tabla de Mermas': {
                'celdas': {
                    'Categorías con medición propia':
                        cm['Categorias con medicion propia'],
                    'Categorías que aún usan la referencia':
                        cm['Categorias con referencia'],
                    'Desviación media de tus mediciones sobre la referencia':
                        cm['Desviacion media de tus mediciones'],
                    'Merma que usas en pescado entero':
                        'G' + str(_fila_merma('Pescado entero')),
                    'Origen del dato en pescado entero':
                        'H' + str(_fila_merma('Pescado entero')),
                    'Merma que usas en verduras de hoja':
                        'G' + str(_fila_merma('Verduras de hoja')),
                    'Origen del dato en verduras de hoja':
                        'H' + str(_fila_merma('Verduras de hoja')),
                    'Merma que usas en aves enteras':
                        'G' + str(_fila_merma('Aves enteras')),
                },
                'tablas': [
                    {'titulo': 'Tabla de mermas de referencia y propias',
                     'cols': cols_mapa(COLS_M),
                     'filas': [M_INI, M_FIN]},
                ],
            },
        },
    }
    destino = os.path.join(AQUI, 'build', 'mapa-' + NOMBRE + '.json')
    with io.open(destino, 'w', encoding='utf-8') as fh:
        fh.write(json.dumps(mapa, ensure_ascii=False, indent=2))
        fh.write(u'\n')
    return destino


def main():
    motor.REGISTRO = []
    wb = Workbook()
    wb.remove(wb.active)
    cr = hoja_rendimiento(wb)
    cc = hoja_coccion(wb)
    cm = hoja_mermas(wb)
    ci = hoja_instrucciones(wb)
    wb.move_sheet('Instrucciones', offset=-3)
    wb.active = 0

    motor.normalizar_texto(wb, NOMBRE + '.xlsx')
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · Versión 1.0 · septiembre 2026'

    destino = os.path.join(AQUI, 'build', NOMBRE + '.xlsx')
    if not os.path.isdir(os.path.dirname(destino)):
        os.makedirs(os.path.dirname(destino))
    wb.save(destino)
    mapa = escribir_mapa(cr, cc, cm, ci)
    print('escrito: ' + destino)
    print('mapa:    ' + mapa)
    print('formulas registradas: ' + str(len(motor.REGISTRO)))
    return destino


if __name__ == '__main__':
    main()
