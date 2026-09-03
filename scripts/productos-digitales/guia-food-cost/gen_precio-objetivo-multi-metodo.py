#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_precio-objetivo-multi-metodo.py — libro 3 de los 8 del producto «Guía
Food Cost + Ingeniería de Menú» (SPEC §2.2, fila 3).

Los cuatro métodos de fijación de precio, plato a plato y a la vez: factor
sobre el coste, margen objetivo en euros, precio de mercado y valor percibido.
El libro calcula el food cost RESULTANTE de cada uno para que se vea dónde el
factor arruina el plato de coste alto.

Cifras: TODAS de `datos_ejemplo.PLATOS` (20 platos; el coste de P1 lo deriva la
ficha). Las siembras de margen, mercado y valor son transformaciones
declaradas de esas cifras, no números nuevos.

Sale SIEMPRE en `build/precio-objetivo-multi-metodo.xlsx`. Después:
    python3 ../inject_cache.py build/precio-objetivo-multi-metodo.xlsx
"""
import io
import json
import os
import sys

AQUI = os.path.dirname(os.path.abspath(__file__))
GUIAS = '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0'
sys.path.insert(0, GUIAS)
sys.path.insert(0, AQUI)

from openpyxl import Workbook                                    # noqa: E402
from openpyxl.formatting.rule import FormulaRule                 # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill         # noqa: E402
from openpyxl.utils import (column_index_from_string,            # noqa: E402
                            get_column_letter)
from openpyxl.worksheet.page import PageMargins                  # noqa: E402
from openpyxl.worksheet.properties import PageSetupProperties    # noqa: E402

import motor                                                     # noqa: E402
import datos_ejemplo as D                                        # noqa: E402

PID = 'guia-food-cost-ingenieria-menu'
PRODUCTO = 'Guía Food Cost + Ingeniería de Menú'
NOMBRE = 'precio-objetivo-multi-metodo'
TITULO = 'Precio Objetivo Multi-Método'
SUBTITULO = 'AI Chef Pro · aichef.pro — Guía Food Cost + Ingeniería de Menú'
VERSION = ('Versión 1.0 · septiembre 2026 · aichef.pro/'
           + PID + ' · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 '
       '· johnguerrero.es')
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde, '
               'desprotege la hoja (sin contraseña).')
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

motor.CTX['producto'] = PID

CAB = '2D2D2D'
CREMA = 'FFF8DC'
FMT_EUR = motor.FMT_EUR
FMT_PCT = motor.FMT_PCT
FMT_ENT = '#,##0'
TIPO_POR_FMT = {FMT_EUR: 'eur', FMT_PCT: 'pct1', FMT_ENT: 'num', None: 'txt'}

# Los cuatro métodos. El texto es el que compara la fórmula del PVP elegido:
# si se toca aquí, hay que tocarlo en `formula_pvp_elegido()`.
METODOS = ['A · Factor sobre el coste', 'B · Margen objetivo',
           'C · Precio de mercado', 'D · Valor percibido']
FAMILIAS = ['Entrantes', 'Principales', 'Postres', 'Menú', 'Bebidas']
DENTRO, FUERA = 'Dentro del objetivo', 'Por encima del objetivo'

# Siembras: transformaciones declaradas del juego de datos único.
# Margen objetivo (€) POR FAMILIA: el MC medio ponderado que calcula
# `matriz-multimetodo-carta!Datos!G40:G42` sobre datos_ejemplo.PLATOS
# (Entrantes 7,46 € · Principales 10,80 € · Postres 4,48 €), redondeado a un
# número manejable para fijar precio. Es UN margen EN EUROS por familia, no
# un factor sobre el coste: por eso el chuletón (coste 14,80 €) no dispara a
# 47 € sino que sale a 25,30 € (14,80 + 10,50), con food cost por encima del
# objetivo — que es justo lo que enseña este método cuando el coste es alto.
MARGEN_OBJETIVO_FAMILIA = {'Entrantes': 6.50, 'Principales': 10.50,
                           'Postres': 4.20}
COSTE_ALTO = 6.0         # € por ración: por encima, el factor deja de servir
COSTE_BAJO = 2.0         # € por ración: por debajo, manda el valor percibido

# Dispersión real por plato para «precio de mercado» (método C) y «valor
# percibido» (método D): cada plato se posiciona distinto frente a su
# competencia, no un único factor sobre el PVP actual (±5-25 % según el
# posicionamiento del plato: los de firma se separan más, los de uso diario
# menos). Clave = id de `datos_ejemplo.PLATOS`.
DISPERSION_MERCADO_VALOR = {
    #        mercado (C)  valor percibido (D)
    'E1':  (1.06, 1.12), 'E2': (0.95, 1.07), 'E3': (1.12, 1.20),
    'E4':  (0.94, 1.08), 'E5': (1.15, 1.22), 'E6': (0.93, 1.06),
    'E7':  (0.90, 1.05), 'P1': (1.08, 1.16), 'P2': (1.12, 1.20),
    'P3':  (0.95, 1.06), 'P4': (1.06, 1.14), 'P5': (1.20, 1.25),
    'P6':  (1.10, 1.18), 'P7': (0.94, 1.07), 'P8': (0.92, 1.05),
    'P9':  (1.15, 1.22), 'D1': (1.06, 1.12), 'D2': (0.94, 1.06),
    'D3':  (1.08, 1.15), 'D4': (0.90, 1.05),
}

F_PAR = 4                # banda de parámetros
F_FC_GLOBAL = 5
F_IVA = 6
F_CAB = 8
F_INI = 9
N_PLATOS = 20                          # platos sembrados de datos_ejemplo
N_FILAS = 25                           # filas totales: 20 sembradas + 5 libres
F_FIN = F_INI + N_FILAS - 1            # 33
F_RES = F_FIN + 2                      # 35

COLS = [
    ('A', '#', 5, FMT_ENT, False),
    ('B', 'Plato', 40, None, True),
    ('C', 'Familia', 14, None, True),
    ('D', 'Coste por ración (€)', 13, FMT_EUR, True),
    ('E', 'FC objetivo del plato (%) — opcional', 13, FMT_PCT, True),
    ('F', 'Método elegido', 18, None, True),
    ('G', 'Margen objetivo (€) — método B', 13, FMT_EUR, True),
    ('H', 'Precio de mercado de la zona (€) — método C', 13, FMT_EUR, True),
    ('I', 'Precio de valor percibido (€) — método D', 13, FMT_EUR, True),
    ('J', 'FC objetivo aplicado (%)', 12, FMT_PCT, False),
    ('K', 'A · PVP por factor (€)', 12, FMT_EUR, False),
    ('L', 'B · PVP por margen (€)', 12, FMT_EUR, False),
    ('M', 'FC resultante con A (%)', 11, FMT_PCT, False),
    ('N', 'FC resultante con B (%)', 11, FMT_PCT, False),
    ('O', 'FC resultante con C (%)', 11, FMT_PCT, False),
    ('P', 'FC resultante con D (%)', 11, FMT_PCT, False),
    ('Q', 'PVP ELEGIDO sin IVA (€)', 14, FMT_EUR, False),
    ('R', 'PVP elegido CON IVA (€)', 14, FMT_EUR, False),
    ('S', 'Margen de contribución (€)', 13, FMT_EUR, False),
    ('T', 'Food cost final (%)', 11, FMT_PCT, False),
    ('U', 'Semáforo vs objetivo', 21, None, False),
    ('V', 'PVP actual en carta sin IVA (€)', 13, FMT_EUR, True),
    ('W', 'Diferencia con el PVP actual (€)', 13, FMT_EUR, False),
]


def metodo_de(coste, familia):
    """Método SEMBRADO por plato, con una regla explícita (no una lista a mano).

    El plato caro no puede ir por factor (el PVP se dispara) y el plato barato
    no puede ir por factor tampoco (el PVP se queda por debajo de lo que el
    cliente pagaría): eso es justo lo que enseña el capítulo 09.
    """
    if coste >= COSTE_ALTO:
        return METODOS[1]
    if coste <= COSTE_BAJO:
        return METODOS[3]
    if familia == 'Principales':
        return METODOS[2]
    return METODOS[0]


def formula_pvp_elegido(r):
    """Cuatro IF encadenados: sin matrices literales, sin CHOOSE, sin INDEX."""
    return ('=IF($F{r}="","",IF($F{r}="{a}",$K{r},IF($F{r}="{b}",$L{r},'
            'IF($F{r}="{c}",$H{r},IF($F{r}="{d}",$I{r},"")))))'
            .format(r=r, a=METODOS[0], b=METODOS[1], c=METODOS[2],
                    d=METODOS[3]))


# --------------------------------------------------------------------------
def a4(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.43, right=0.43, top=0.55, bottom=0.55,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def _letras(desde, hasta):
    return [get_column_letter(i)
            for i in range(column_index_from_string(desde),
                           column_index_from_string(hasta) + 1)]


def banda(ws, fila, desde, hasta, texto):
    motor.val(ws, desde + str(fila), texto, bold=True)
    for col in _letras(desde, hasta):
        ws[col + str(fila)].fill = PatternFill('solid', fgColor=CREMA)
        ws[col + str(fila)].font = Font(bold=True, size=11)


def semaforo_doble(ws, rango, ancla, umbral):
    motor.regla_expresion(
        ws, rango,
        '=AND(ISNUMBER(' + ancla + '),' + ancla + '>' + umbral + ')',
        bg=motor.CF_ROJO_BG, fg=motor.CF_ROJO_FG)
    ws.conditional_formatting.add(
        rango,
        FormulaRule(
            formula=['=AND(ISNUMBER(' + ancla + '),' + ancla + '<=' + umbral
                     + ')'],
            stopIfTrue=True,
            font=Font(color=motor.CF_VERDE_FG, bold=True),
            fill=PatternFill(start_color=motor.CF_VERDE_BG,
                             end_color=motor.CF_VERDE_BG, fill_type='solid')))


# --------------------------------------------------------------------------
def hoja_por_plato(wb):
    ws = wb.create_sheet('Por Plato')
    motor.val(ws, 'A1', TITULO, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color='595959')
    motor.val(ws, 'B3', 'El mismo plato con los cuatro métodos a la vez. Elige '
                        'el método en la columna F y el libro se queda con ese '
                        'precio.', wrap=True)
    motor.anchos(ws, dict((c[0], c[2]) for c in COLS))

    # ---- parámetros ------------------------------------------------------
    banda(ws, F_PAR, 'B', 'I',
          'PARÁMETROS — cámbialos y se recalcula todo el libro')
    motor.val(ws, 'B' + str(F_FC_GLOBAL), 'Food cost objetivo global (%)',
              bold=True)
    motor.val(ws, 'E' + str(F_FC_GLOBAL),
              D.RESTAURANTE['food_cost_objetivo'], fmt=FMT_PCT)
    motor.verde(ws, 'E' + str(F_FC_GLOBAL))
    motor.val(ws, 'J' + str(F_FC_GLOBAL),
              'Se aplica a todos los platos salvo a los que tengan su propio '
              'objetivo en la columna E.', wrap=True)
    motor.escribir_parametro(ws, F_IVA, 'B', 'E', 'iva_restauracion',
                             col_nota='J')

    # ---- cabecera y datos ------------------------------------------------
    for letra, texto, _, _, _ in COLS:
        cel = motor.val(ws, letra + str(F_CAB), texto, bold=True, wrap=True)
        cel.fill = PatternFill('solid', fgColor=CAB)
        cel.font = Font(bold=True, color='FFFFFF', size=9)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    ws.row_dimensions[F_CAB].height = 52
    ws.freeze_panes = 'C' + str(F_INI)
    ws.auto_filter.ref = 'A' + str(F_CAB) + ':W' + str(F_FIN)

    for i in range(N_FILAS):
        r = F_INI + i
        motor.val(ws, 'A' + str(r), i + 1, fmt=FMT_ENT, align='center')
        if i < N_PLATOS:
            pid_plato, nombre, familia, coste, pvp, _uds = D.PLATOS[i]
            motor.val(ws, 'B' + str(r), nombre)
            motor.val(ws, 'C' + str(r), familia)
            motor.val(ws, 'D' + str(r), coste, fmt=FMT_EUR)
            ws['E' + str(r)].number_format = FMT_PCT
            motor.val(ws, 'F' + str(r), metodo_de(coste, familia))
            motor.val(ws, 'G' + str(r), MARGEN_OBJETIVO_FAMILIA[familia],
                      fmt=FMT_EUR)
            f_mercado, f_valor = DISPERSION_MERCADO_VALOR[pid_plato]
            motor.val(ws, 'H' + str(r), round(pvp * f_mercado, 2),
                      fmt=FMT_EUR)
            motor.val(ws, 'I' + str(r), round(pvp * f_valor, 2), fmt=FMT_EUR)
            motor.val(ws, 'V' + str(r), pvp, fmt=FMT_EUR)
        else:
            ws['D' + str(r)].number_format = FMT_EUR
            ws['E' + str(r)].number_format = FMT_PCT
            ws['G' + str(r)].number_format = FMT_EUR
            ws['H' + str(r)].number_format = FMT_EUR
            ws['I' + str(r)].number_format = FMT_EUR
            ws['V' + str(r)].number_format = FMT_EUR
        for letra, _, _, _, entrada in COLS:
            if entrada:
                motor.verde(ws, letra + str(r))

        motor.f(ws, 'J' + str(r),
                '=IF(ISNUMBER($E{r}),$E{r},IF($E${g}="","",$E${g}))'
                .format(r=r, g=F_FC_GLOBAL), fmt=FMT_PCT)
        motor.f(ws, 'K' + str(r),
                '=IFERROR(IF(OR($D{r}="",$J{r}="",$J{r}=0),"",$D{r}/$J{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'L' + str(r),
                '=IFERROR(IF(OR($D{r}="",$G{r}=""),"",$D{r}+$G{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        for col_fc, col_pvp in (('M', 'K'), ('N', 'L'), ('O', 'H'),
                                ('P', 'I')):
            motor.f(ws, col_fc + str(r),
                    '=IFERROR(IF(OR($D{r}="",${p}{r}="",${p}{r}=0),"",'
                    '$D{r}/${p}{r}),"")'.format(r=r, p=col_pvp), fmt=FMT_PCT)
        motor.f(ws, 'Q' + str(r), formula_pvp_elegido(r), fmt=FMT_EUR,
                bold=True)
        motor.f(ws, 'R' + str(r),
                '=IFERROR(IF($Q{r}="","",$Q{r}*(1+$E${i})),"")'
                .format(r=r, i=F_IVA), fmt=FMT_EUR)
        motor.f(ws, 'S' + str(r),
                '=IFERROR(IF(OR($Q{r}="",$D{r}=""),"",$Q{r}-$D{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'T' + str(r),
                '=IFERROR(IF(OR($D{r}="",$Q{r}="",$Q{r}=0),"",$D{r}/$Q{r}),"")'
                .format(r=r), fmt=FMT_PCT, bold=True)
        motor.f(ws, 'U' + str(r),
                '=IF(OR($T{r}="",$J{r}=""),"",IF($T{r}<=$J{r},"{d}","{f}"))'
                .format(r=r, d=DENTRO, f=FUERA))
        motor.f(ws, 'W' + str(r),
                '=IFERROR(IF(OR($Q{r}="",$V{r}=""),"",$Q{r}-$V{r}),"")'
                .format(r=r), fmt=FMT_EUR)

    rango_t = 'T' + str(F_INI) + ':T' + str(F_FIN)
    semaforo_doble(ws, rango_t, '$T' + str(F_INI), '$J' + str(F_INI))
    motor.semaforo_texto(ws, 'U' + str(F_INI) + ':U' + str(F_FIN),
                         ((DENTRO, motor.CF_VERDE_BG, motor.CF_VERDE_FG),
                          (FUERA, motor.CF_ROJO_BG, motor.CF_ROJO_FG)))

    # ---- resumen ---------------------------------------------------------
    a, b = F_INI, F_FIN
    r = F_RES
    banda(ws, r, 'B', 'I', 'RESUMEN — lo calcula el libro')
    celdas = {}

    r += 1
    celdas['Platos con precio calculado'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Platos con precio calculado (n.º)')
    motor.f(ws, 'E' + str(r),
            '=COUNTIF($Q${a}:$Q${b},">0")'.format(a=a, b=b), fmt=FMT_ENT)

    r += 1
    celdas['Platos dentro del objetivo'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Platos dentro del objetivo (n.º)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=COUNTIF($U${a}:$U${b},"{d}")'.format(a=a, b=b, d=DENTRO),
            fmt=FMT_ENT, bold=True)

    r += 1
    celdas['Platos por encima del objetivo'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r), 'Platos por encima del objetivo (n.º)',
              bold=True)
    motor.f(ws, 'E' + str(r),
            '=COUNTIF($U${a}:$U${b},"{f}")'.format(a=a, b=b, f=FUERA),
            fmt=FMT_ENT, bold=True)

    r += 1
    celdas['Food cost del conjunto de la carta'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Food cost del conjunto (suma de costes ÷ suma de PVP '
              'elegidos) (%)', bold=True)
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(SUM($Q${a}:$Q${b})=0,"",'
            'SUMIF($Q${a}:$Q${b},">0",$D${a}:$D${b})/SUM($Q${a}:$Q${b})),"")'
            .format(a=a, b=b), fmt=FMT_PCT, bold=True)
    motor.val(ws, 'J' + str(r),
              'No está ponderado por ventas: para eso está la matriz de '
              'ingeniería de menú.', wrap=True)

    r += 1
    celdas['Margen de contribución medio por plato'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Margen de contribución medio por plato (€)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(COUNTIF($Q${a}:$Q${b},">0")=0,"",'
            'SUM($S${a}:$S${b})/COUNTIF($Q${a}:$Q${b},">0")),"")'
            .format(a=a, b=b), fmt=FMT_EUR)

    r += 1
    celdas['Diferencia total con los PVP actuales'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Diferencia total con los PVP actuales de la carta (€)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(COUNTIF($Q${a}:$Q${b},">0")=0,"",'
            'SUM($W${a}:$W${b})),"")'.format(a=a, b=b), fmt=FMT_EUR)

    r += 1
    celdas['Subida media sobre el PVP actual'] = 'E' + str(r)
    motor.val(ws, 'B' + str(r),
              'Subida media sobre el PVP actual de la carta (%)')
    motor.f(ws, 'E' + str(r),
            '=IFERROR(IF(SUMIF($Q${a}:$Q${b},">0",$V${a}:$V${b})=0,"",'
            'SUM($W${a}:$W${b})/SUMIF($Q${a}:$Q${b},">0",$V${a}:$V${b})),"")'
            .format(a=a, b=b), fmt=FMT_PCT)

    r += 2
    banda(ws, r, 'B', 'I',
          'LA CARTA ENTERA POR CADA MÉTODO — para ver de qué hablamos')
    f_metodos = r + 1
    for etiqueta, col, clave in (
            ('Facturación de la carta si TODA fuese por factor (€)', 'K',
             'Carta entera por factor'),
            ('Facturación si TODA fuese por margen objetivo (€)', 'L',
             'Carta entera por margen'),
            ('Facturación si TODA fuese a precio de mercado (€)', 'H',
             'Carta entera a mercado'),
            ('Facturación si TODA fuese a valor percibido (€)', 'I',
             'Carta entera a valor'),
            ('Facturación con los PVP actuales de la carta (€)', 'V',
             'Carta entera con los PVP actuales')):
        r += 1
        celdas[clave] = 'E' + str(r)
        motor.val(ws, 'B' + str(r), etiqueta)
        motor.f(ws, 'E' + str(r),
                '=IF(COUNTIF(${c}${a}:${c}${b},">0")=0,"",'
                'SUM(${c}${a}:${c}${b}))'.format(c=col, a=a, b=b), fmt=FMT_EUR)
    motor.val(ws, 'J' + str(f_metodos),
              'Suma de un PVP por plato, sin ponderar por unidades vendidas: '
              'es un orden de magnitud para comparar métodos, no una previsión '
              'de ventas.', wrap=True)

    r += 2
    banda(ws, r, 'B', 'I', 'REPARTO DE MÉTODOS EN TU CARTA')
    for m in METODOS:
        r += 1
        celdas['Platos con el método ' + m[0]] = 'E' + str(r)
        motor.val(ws, 'B' + str(r), 'Platos con el método ' + m + ' (n.º)')
        motor.f(ws, 'E' + str(r),
                '=COUNTIF($F${a}:$F${b},"{m}")'.format(a=a, b=b, m=m),
                fmt=FMT_ENT)

    # ---- validación de datos --------------------------------------------
    filas = [str(x) for x in range(F_INI, F_FIN + 1)]
    motor.dv_lista(ws, ['C' + x for x in filas], FAMILIAS,
                   titulo='Familia no válida')
    motor.dv_lista(ws, ['F' + x for x in filas], METODOS,
                   titulo='Método no válido',
                   mensaje='Elige uno de los cuatro métodos de la lista.')
    motor.dv_numerica(ws, ['D' + x for x in filas], minimo=0,
                      titulo='Coste por ración',
                      prompt='El coste por ración de la ficha de escandallo, '
                             'sin IVA.')
    motor.dv_porcentaje(ws, ['E' + x for x in filas],
                        titulo='FC objetivo del plato',
                        prompt='Déjalo vacío para usar el objetivo global. '
                               'Se escribe en tanto por uno: 0,45 = 45 %.')
    motor.dv_numerica(ws, ['G' + x for x in filas], minimo=0,
                      titulo='Margen objetivo',
                      prompt='Euros de margen que quieres dejar en el plato '
                             '(método B).')
    motor.dv_numerica(ws, ['H' + x for x in filas], minimo=0,
                      titulo='Precio de mercado',
                      prompt='Lo que cobran por ese plato a tu alrededor, sin '
                             'IVA (método C).')
    motor.dv_numerica(ws, ['I' + x for x in filas], minimo=0,
                      titulo='Precio de valor percibido',
                      prompt='Lo que tu cliente aceptaría pagar, sin IVA '
                             '(método D).')
    motor.dv_numerica(ws, ['V' + x for x in filas], minimo=0,
                      titulo='PVP actual',
                      prompt='El precio que hoy está en tu carta, sin IVA.')
    motor.dv_porcentaje(ws, ['E' + str(F_FC_GLOBAL)],
                        titulo='Food cost objetivo',
                        prompt='Se escribe en tanto por uno: 0,30 = 30 %.')
    motor.dv_porcentaje(ws, ['E' + str(F_IVA)], titulo='Tipo de IVA',
                        maximo=0.21,
                        prompt='Tipo de IVA repercutido del canal. 0,10 en '
                               'sala.')

    a4(ws, apaisado=True, titulos=str(F_CAB) + ':' + str(F_CAB))
    return celdas


# --------------------------------------------------------------------------
PASOS = [
    '1. Trae de la ficha de escandallo el coste por ración de cada plato a la '
    'columna D. Es el único dato que no se negocia.',
    '2. Rellena las tres entradas de los métodos: margen objetivo en euros '
    '(B), precio de mercado de tu zona (C) y precio de valor percibido (D).',
    '3. Mira las cuatro columnas «FC resultante»: el mismo plato con cuatro '
    'precios distintos y cuatro food cost distintos.',
    '4. Elige el método en la columna «Método elegido» y el libro se queda con '
    'ese precio, le pone el IVA y calcula el margen.',
    '5. El semáforo compara el food cost final con tu objetivo: verde si lo '
    'cumples, rojo si te pasas.',
    '6. Deja vacía la columna «FC objetivo del plato» para usar el objetivo '
    'global; escribe un objetivo propio en los platos de coste alto.',
    '7. La columna «Diferencia con el PVP actual» es la subida (o la bajada) '
    'que estás proponiendo: llévala al plan de 90 días antes de tocar la carta.',
    '8. La tabla trae los 20 platos del ejemplo y 5 filas libres ya '
    'formateadas y validadas: añade tus platos ahí antes de insertar filas '
    'nuevas.',
]

NOTA_IVA_LIBRO = (
    'El tipo de IVA de este libro es el de SALA: 10 % para la comida y la '
    'bebida consumidas en el local, alcohol incluido (art. 91.Uno.2.2.º de la '
    'Ley 37/1992). Para llevar y en delivery el tipo depende del producto: '
    'cámbialo en la celda verde «Tipo de IVA de restauración (%)».')

NOTA_METODOS = (
    'Los métodos sembrados en el ejemplo siguen una regla: coste por ración de '
    '6 € o más, margen objetivo; 2 € o menos, valor percibido; el resto de '
    'principales, precio de mercado; y los demás, factor sobre el coste. El '
    'factor solo funciona en la banda intermedia: en el chuletón dispara el '
    'precio y en el postre lo deja por debajo de lo que el cliente pagaría. '
    'El margen objetivo (método B) va sembrado en EUROS POR FAMILIA (6,50 € '
    'entrantes, 10,50 € principales, 4,20 € postres) — el MC medio ponderado '
    'de la carta, no un múltiplo del coste: por eso el chuletón sale a '
    '25,30 € (14,80 + 10,50) y no a 47 €, y su food cost queda por encima '
    'del objetivo en vez de disparado. El precio de mercado y el valor '
    'percibido llevan dispersión propia por plato (entre el 5 % y el 25 % '
    'sobre el PVP actual, según lo diferenciado que está cada plato de su '
    'competencia), no un único factor para toda la carta.')


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 112
    motor.val(ws, 'A1', TITULO, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color='595959')
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), 'Por qué cada plato lleva un método',
              bold=True)
    fila += 1
    coords = {'nota_metodos': 'A' + str(fila)}
    motor.val(ws, 'A' + str(fila), NOTA_METODOS, wrap=True)
    fila += 2
    for texto in (motor.NOTA_VERDES, NOTA_IVA_LIBRO, DESPROTEGER, BIO,
                  VERSION):
        cel = motor.val(ws, 'A' + str(fila), texto, wrap=True)
        if texto is motor.NOTA_VERDES:
            cel.fill = PatternFill('solid', fgColor=motor.VERDE)
        if texto is VERSION:
            coords['version'] = 'A' + str(fila)
        fila += 1
    a4(ws, apaisado=False)
    return coords


# --------------------------------------------------------------------------
def _fila_plato(pid_plato):
    for i, p in enumerate(D.PLATOS):
        if p[0] == pid_plato:
            return F_INI + i
    raise KeyError(pid_plato)


def escribir_mapa(cp, ci):
    chuleton, croquetas = _fila_plato('P5'), _fila_plato('E1')
    celdas = {'Food cost objetivo global': 'E' + str(F_FC_GLOBAL),
              'Tipo de IVA de restauración (sala)': 'E' + str(F_IVA),
              # Los dos extremos que sostienen el capítulo 09: el plato caro
              # que el factor dispara y el barato que el factor hunde.
              'PVP del chuletón por factor': 'K' + str(chuleton),
              'PVP del chuletón por margen objetivo': 'L' + str(chuleton),
              'PVP elegido del chuletón': 'Q' + str(chuleton),
              'Food cost final del chuletón': 'T' + str(chuleton),
              'PVP actual del chuletón en carta': 'V' + str(chuleton),
              'PVP de las croquetas por factor': 'K' + str(croquetas),
              'PVP elegido de las croquetas': 'Q' + str(croquetas),
              'PVP actual de las croquetas en carta': 'V' + str(croquetas),
              'Diferencia de las croquetas con el PVP actual':
                  'W' + str(croquetas)}
    celdas.update(cp)
    mapa = {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Instrucciones': {
                'celdas': {'Título del libro': 'A1',
                           'Regla de asignación de métodos':
                               ci['nota_metodos'],
                           'Línea de versión': ci['version']},
                'tablas': [],
            },
            'Por Plato': {
                'celdas': celdas,
                # Solo rangos HOMOGÉNEOS con cabecera real. «La carta entera
                # por cada método» y «Reparto de métodos» son bloques
                # etiqueta/valor (la columna B lleva la frase completa, no
                # una cabecera de tabla «Método»/«Facturación»): sus celdas
                # ya están en «celdas» arriba (vía `cp`, con las claves
                # cortas «Carta entera por factor», «Platos con el método
                # A»…), y por eso NO se declaran aquí como tabla — un
                # capítulo que copiase esta cabecera imprimiría un rótulo
                # que no existe en el xlsx.
                'tablas': [
                    {'titulo': 'Precio por los cuatro métodos, plato a plato',
                     'cols': [[c[1], c[0], TIPO_POR_FMT[c[3]]] for c in COLS],
                     'filas': [F_INI, F_FIN]},
                ],
            },
        },
    }
    destino = os.path.join(AQUI, 'build', 'mapa-' + NOMBRE + '.json')
    with io.open(destino, 'w', encoding='utf-8') as fh:
        fh.write(json.dumps(mapa, ensure_ascii=False, indent=2))
        fh.write(u'\n')
    return destino


def main():
    motor.REGISTRO = []
    wb = Workbook()
    wb.remove(wb.active)
    cp = hoja_por_plato(wb)
    ci = hoja_instrucciones(wb)
    wb.move_sheet('Instrucciones', offset=-1)
    wb.active = 0

    motor.normalizar_texto(wb, NOMBRE + '.xlsx')
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · Versión 1.0 · septiembre 2026'

    destino = os.path.join(AQUI, 'build', NOMBRE + '.xlsx')
    if not os.path.isdir(os.path.dirname(destino)):
        os.makedirs(os.path.dirname(destino))
    wb.save(destino)
    mapa = escribir_mapa(cp, ci)
    print('escrito: ' + destino)
    print('mapa:    ' + mapa)
    print('formulas registradas: ' + str(len(motor.REGISTRO)))
    return destino


if __name__ == '__main__':
    main()
