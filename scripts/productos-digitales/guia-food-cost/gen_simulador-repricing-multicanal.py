#!/usr/bin/env python3
"""
gen_simulador-repricing-multicanal.py — libro 5 de «Guía Food Cost + Ingeniería
de Menú» (SPEC §2.2 fila 5, decisiones D4 y D15).

Hojas: Instrucciones · Parámetros · Carta · Resumen.

Qué responde el libro: si llevas la carta del comedor a take away y a delivery
SIN tocar el precio, ¿en qué se convierte tu food cost? ¿Qué precio necesitas
en cada canal para mantener el objetivo? ¿Cabe ese precio bajo el techo que
acepta el mercado? ¿Y cuánto margen deja cada canal con las mismas unidades?

Convenciones de la familia (motor.py de guias-v2_0):
  · celdas verdes = entrada; ninguna celda con fórmula va en verde;
  · cero constantes tecleadas dentro de una fórmula: el IVA, la comisión, el
    packaging y el food cost objetivo viven en celdas y las fórmulas los leen
    con referencias absolutas;
  · «sin dato» = "" (nunca 0); IFERROR en todo cociente; semáforos con ISNUMBER;
  · funciones PROHIBIDAS (pycel + Google Sheets/Numbers): INDIRECT, COUNTA,
    PMT, OFFSET, XLOOKUP, LET, LAMBDA y matrices dinámicas.

Truco de compatibilidad: los parámetros de cada canal se ESPEJAN en la propia
hoja «Carta» (bloque «PARÁMETROS EN VIGOR», filas 27-31) con fórmulas que leen
«Parámetros». Así las fórmulas de fila y —sobre todo— el formato condicional
no cruzan de hoja: Google Sheets no admite referencias a otra hoja dentro de
una regla de formato condicional y la única forma de hacerlo sería INDIRECT,
que está prohibida.

Salida: build/simulador-repricing-multicanal.xlsx (+ mapa en el informe).
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
sys.path.insert(0, AQUI)

import motor                                                   # noqa: E402
import datos_ejemplo as D                                      # noqa: E402

motor.CTX['producto'] = 'guia-food-cost-ingenieria-menu'

# --------------------------------------------------------------------------
# Constantes de producto (NO se usa motor.version_line: esta guía es la 1.0)
# --------------------------------------------------------------------------
PRODUCTO = 'Guía Food Cost + Ingeniería de Menú'
SUBTITULO = 'AI Chef Pro · aichef.pro — ' + PRODUCTO
VERSION = ('Versión 1.0 · septiembre 2026 · '
           'aichef.pro/guia-food-cost-ingenieria-menu · info@aichef.pro')
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010 '
       '· johnguerrero.es')
NOTA_DESPROTEGER = ('Para editar la estructura o una celda que no esté en '
                    'verde, desprotege la hoja (sin contraseña).')
TITULO = 'Simulador de repricing multicanal'
NOMBRE = 'simulador-repricing-multicanal'

FMT_EUR = '#,##0.00 €'
FMT_PCT = '0.0%'
FMT_ENT = '#,##0'
FMT_CANT = '0.000'

GRIS = 'F2F2F2'
CREMA = 'FFF6DC'
ORO = 'FFD700'
CABECERA = '2D2D2D'

CANALES = ('Sala', 'Take away', 'Delivery')
TIPOS = ('Comida', 'Refresco/azucarada', 'Bebida alcohólica')


# --------------------------------------------------------------------------
# Utilidades de estilo (idénticas en los tres libros del constructor C)
# --------------------------------------------------------------------------
def cabecera(ws, fila, columnas, altura=34):
    """Fila de cabecera: fondo 2D2D2D, letra blanca, centrada y con ajuste."""
    for letra, texto in columnas:
        c = ws[letra + str(fila)]
        c.value = texto
        c.fill = PatternFill('solid', fgColor=CABECERA)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    ws.row_dimensions[fila].height = altura


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12)


def fila_total(ws, fila, primera, ultima):
    """Pinta de crema/oro una fila de totales entre dos columnas."""
    for col in range(motor.column_index_from_string(primera),
                     motor.column_index_from_string(ultima) + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = PatternFill('solid', fgColor=CREMA)
        c.font = Font(bold=True)


def cf_expresion(ws, rango, formula, bg, fg):
    """Añade una regla de formato condicional SIN purgar las anteriores."""
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


def cf_texto(ws, rango, vocabulario):
    for texto, bg, fg in vocabulario:
        ws.conditional_formatting.add(
            rango, Rule(type='cellIs', operator='equal',
                        formula=['"' + texto + '"'], stopIfTrue=True,
                        dxf=DifferentialStyle(
                            font=Font(color=fg, bold=True),
                            fill=PatternFill(start_color=bg, end_color=bg,
                                             fill_type='solid'))))


def pagina(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9                     # A4
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezar(ws, titulo, nota=None):
    motor.val(ws, 'A1', titulo)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    if nota:
        motor.val(ws, 'A3', nota)
        ws['A3'].font = Font(italic=True, size=9)


def anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


# --------------------------------------------------------------------------
# Hoja Instrucciones
# --------------------------------------------------------------------------
PASOS = [
    '1. Hoja «Carta»: escribe en las celdas verdes cada plato con su coste por '
    'ración, el PVP que tiene HOY en el comedor (sin IVA), el tipo de producto, '
    'el precio techo que acepta el mercado y las unidades que vendes al mes.',
    '2. Hoja «Parámetros»: pon la comisión, el packaging por pedido, los platos '
    'por pedido y el food cost objetivo de CADA canal. Son los de tu contrato, '
    'no una tarifa oficial.',
    '3. Revisa la matriz de IVA repercutido: manda el canal y el tipo de '
    'producto, y las fórmulas la leen con INDEX/MATCH.',
    '4. Lee «Food cost efectivo al PVP de sala»: es en lo que se convierte tu '
    'food cost si publicas en la app el mismo precio que tienes en el comedor.',
    '5. Compara el «PVP sin IVA necesario» con tu precio techo. La columna '
    '«¿Viable?» marca los platos que no caben en ese canal a tu objetivo.',
    '6. «PVP que puedes cobrar» es el menor de los dos: si el precio necesario '
    'pasa del techo, cobras el techo y el food cost se te va; ahí está el '
    'margen mensual real de cada canal.',
    '7. Hoja «Resumen»: platos viables, food cost medio, margen mensual por '
    'canal y diferencia frente a la sala.',
    '8. Las unidades vendidas son las mismas en los tres canales a propósito: '
    'la pregunta que responde el libro es qué margen daría ESE volumen en cada '
    'canal, no cuánto venderías en cada uno.',
]

NOTAS_LIBRO = [
    'Todos los importes van SIN IVA salvo la columna «PVP con IVA para el '
    'cliente», que es el precio que el comensal ve en la carta o en la app.',
    'El IVA repercutido depende del canal Y del tipo de producto: en sala todo '
    'va al 10 % (art. 91.Uno.2.2.º de la Ley del IVA, alcohol incluido); para '
    'llevar y en reparto es entrega de bienes, así que la comida va al 10 % '
    '(art. 91.Uno.1.1.º) y las bebidas alcohólicas y los refrescos con azúcares '
    'o edulcorantes añadidos, al 21 %.',
    'El food cost se mide sobre la venta NETA (base imponible) y con el coste '
    'NETO de IVA soportado: el IVA de las compras se deduce en el modelo 303, '
    'es tesorería, no coste.',
    'El coste por ración sale de la ficha de escandallo (libro '
    '«ficha-escandallo-base.xlsx» de este mismo pack).',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    anchos(ws, {'A': 110.0})
    motor.val(ws, 'A1', TITULO)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    seccion(ws, 'A4', 'Instrucciones de uso')
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        ws.row_dimensions[fila].height = 28
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), 'Celdas verdes = campos editables')
    ws['A' + str(fila)].fill = PatternFill('solid', fgColor=motor.VERDE)
    fila += 2
    for nota in NOTAS_LIBRO:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        ws.row_dimensions[fila].height = 28
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), NOTA_DESPROTEGER)
    fila += 1
    motor.val(ws, 'A' + str(fila), BIO)
    fila += 1
    motor.val(ws, 'A' + str(fila), VERSION)
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Hoja Parámetros
# --------------------------------------------------------------------------
def hoja_parametros(wb):
    ws = wb.create_sheet('Parámetros')
    anchos(ws, {'A': 24, 'B': 30, 'C': 22, 'D': 20, 'E': 22, 'F': 22})
    encabezar(ws, 'Parámetros del simulador')

    seccion(ws, 'A4', 'MATRIZ DE IVA REPERCUTIDO POR CANAL Y TIPO DE PRODUCTO')
    cabecera(ws, 5, [('A', 'Canal'), ('B', TIPOS[0]), ('C', TIPOS[1]),
                     ('D', TIPOS[2])])
    for i, canal in enumerate(CANALES):
        fila = 6 + i
        motor.val(ws, 'A' + str(fila), canal, bold=True)
        for j, tipo in enumerate(TIPOS):
            col = chr(ord('B') + j)
            motor.val(ws, col + str(fila), D.IVA_REPERCUTIDO[canal][tipo],
                      fmt=FMT_PCT, verde_=True)
    fila = 9
    for canal in CANALES:
        motor.val(ws, 'A' + str(fila), canal + ' — ' + D.NOTAS_IVA[canal])
        fila += 1
    motor.val(ws, 'A12', 'Fuente: ' + D.FUENTE_IVA)

    seccion(ws, 'A14', 'PARÁMETROS DEL SIMULADOR POR CANAL')
    cabecera(ws, 15, [('A', 'Canal'),
                      ('B', 'Comisión de la plataforma (%)'),
                      ('C', 'Packaging (€/pedido)'),
                      ('D', 'Platos por pedido (n.º)'),
                      ('E', 'Food cost objetivo (%)'),
                      ('F', 'Packaging por plato (€)')])
    defecto = {
        'Sala': (0.0, 0.0, 1.0, D.SIMULADOR_DEFECTO['fc_objetivo_sala']),
        'Take away': (D.SIMULADOR_DEFECTO['comision_take_away'],
                      D.SIMULADOR_DEFECTO['packaging_por_pedido'],
                      D.SIMULADOR_DEFECTO['platos_por_pedido'],
                      D.SIMULADOR_DEFECTO['fc_objetivo_take_away']),
        'Delivery': (D.SIMULADOR_DEFECTO['comision_delivery'],
                     D.SIMULADOR_DEFECTO['packaging_por_pedido'],
                     D.SIMULADOR_DEFECTO['platos_por_pedido'],
                     D.SIMULADOR_DEFECTO['fc_objetivo_delivery']),
    }
    for i, canal in enumerate(CANALES):
        fila = 16 + i
        com, pack, platos, obj = defecto[canal]
        motor.val(ws, 'A' + str(fila), canal, bold=True)
        motor.val(ws, 'B' + str(fila), com, fmt=FMT_PCT, verde_=True)
        motor.val(ws, 'C' + str(fila), pack, fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'D' + str(fila), platos, fmt=FMT_CANT, verde_=True)
        motor.val(ws, 'E' + str(fila), obj, fmt=FMT_PCT, verde_=True)
        motor.f(ws, 'F' + str(fila),
                '=IFERROR(IF(OR($C{0}="",$D{0}="",$D{0}=0),"",$C{0}/$D{0}),"")'
                .format(fila), fmt=FMT_EUR)
    motor.val(ws, 'A19',
              'En sala no hay packaging ni comisión: el packaging por plato '
              'sale de dividir el envase de un pedido entre los platos que '
              'lleva ese pedido.')
    motor.val(ws, 'A20',
              'La comisión se aplica sobre el precio CON el que vendes en la '
              'plataforma; el food cost objetivo puede ser distinto en cada '
              'canal si decides que el delivery trabaje con otro margen.')

    seccion(ws, 'A22', 'COMISIONES DE LAS PLATAFORMAS — ORDEN DE MAGNITUD '
                       '(INFORMATIVO, NO ES UN TARIFARIO)')
    cabecera(ws, 23, [('A', 'Plataforma'),
                      ('B', 'Comisión y cuotas de referencia')], altura=20)
    for i, (plataforma, texto) in enumerate(D.COMISIONES_REFERENCIA):
        fila = 24 + i
        motor.val(ws, 'A' + str(fila), plataforma, bold=True)
        motor.val(ws, 'B' + str(fila), texto)
    motor.val(ws, 'A28', 'Fuente: ' + D.FUENTE_COMISIONES)
    rango_pack = '-'.join(('%.2f' % x).replace('.', ',')
                          for x in D.PACKAGING_REFERENCIA)
    motor.val(ws, 'A29',
              'Packaging de referencia: ' + rango_pack + ' €/pedido (misma '
              'fuente). Usa el coste real de TU envase.')

    motor.dv_porcentaje(ws, ['B%d' % r for r in (6, 7, 8)]
                        + ['C%d' % r for r in (6, 7, 8)]
                        + ['D%d' % r for r in (6, 7, 8)],
                        titulo='Tipo de IVA',
                        prompt='Se escribe en tanto por uno: 0,10 = 10 %.')
    motor.dv_porcentaje(ws, ['B16', 'B17', 'B18'], titulo='Comisión',
                        prompt='Se escribe en tanto por uno: 0,30 = 30 %.',
                        maximo=0.95)
    motor.dv_porcentaje(ws, ['E16', 'E17', 'E18'], titulo='Food cost objetivo',
                        prompt='Se escribe en tanto por uno: 0,30 = 30 %.',
                        maximo=0.95)
    motor.dv_numerica(ws, ['C16', 'C17', 'C18'], minimo=0,
                      titulo='Packaging (€/pedido)')
    motor.dv_numerica(ws, ['D16', 'D17', 'D18'], minimo=0.1, maximo=50,
                      titulo='Platos por pedido')
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Hoja Carta
# --------------------------------------------------------------------------
FILA_CAB = 4
FILA_INI = 5
CAP_PLATOS = 25            # 20 platos sembrados + 5 filas libres (M2)
FILA_FIN = FILA_INI + CAP_PLATOS - 1
FILA_TOT = FILA_FIN + 1
FILA_MIRROR = FILA_TOT + 4  # Sala; +1 Take away; +2 Delivery

BLOQUES = (('Sala', 'H'), ('Take away', 'O'), ('Delivery', 'V'))


def _col(letra, salto):
    return motor.get_column_letter(motor.column_index_from_string(letra)
                                   + salto)


def hoja_carta(wb):
    ws = wb.create_sheet('Carta')
    encabezar(ws, 'Carta: el mismo plato en sala, take away y delivery',
              nota='El coste por ración sale de la ficha de escandallo; el PVP '
                   'de sala es el que tienes hoy en la carta, sin IVA.')
    base = [('A', '#'), ('B', 'Plato'),
            ('C', 'Coste por ración, sin IVA (€)'),
            ('D', 'PVP en sala, sin IVA (€)'),
            ('E', 'Tipo de producto (IVA)'),
            ('F', 'Precio techo del mercado, sin IVA (€)'),
            ('G', 'Uds vendidas al mes (n.º)')]
    cabecera(ws, FILA_CAB, base, altura=48)
    anchos(ws, {'A': 5, 'B': 40, 'C': 13, 'D': 13, 'E': 17, 'F': 15, 'G': 12})

    cols_bloque = ('Food cost efectivo al PVP de sala (%)',
                   'PVP sin IVA necesario para el food cost objetivo (€)',
                   'PVP con IVA para el cliente (€)',
                   '¿Viable? (PVP necesario vs precio techo)',
                   'PVP que puedes cobrar, sin IVA (€)',
                   'Ingreso neto por plato (€)',
                   'Margen mensual (€)')
    for canal, ini in BLOQUES:
        motor.val(ws, ini + '3', canal, bold=True)
        ws[ini + '3'].fill = PatternFill('solid', fgColor=CREMA)
        ws[ini + '3'].alignment = Alignment(horizontal='center')
        cabecera(ws, FILA_CAB,
                 [(_col(ini, i), t) for i, t in enumerate(cols_bloque)],
                 altura=48)
        for i, ancho in enumerate((14, 16, 15, 20, 15, 15, 15)):
            ws.column_dimensions[_col(ini, i)].width = ancho

    for i in range(CAP_PLATOS):
        fila = FILA_INI + i
        motor.val(ws, 'A' + str(fila), i + 1, fmt=FMT_ENT)
        if i < len(D.PLATOS):
            pid, nombre, familia, coste, pvp, uds = D.PLATOS[i]
            motor.val(ws, 'B' + str(fila), nombre, verde_=True)
            motor.val(ws, 'C' + str(fila), coste, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'D' + str(fila), pvp, fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'E' + str(fila), D.TIPO_PRODUCTO_PLATO, verde_=True)
            motor.val(ws, 'F' + str(fila), D.PRECIO_TECHO_APP[pid],
                      fmt=FMT_EUR, verde_=True)
            motor.val(ws, 'G' + str(fila), uds, fmt=FMT_ENT, verde_=True)
        else:
            ws['C' + str(fila)].number_format = FMT_EUR
            ws['D' + str(fila)].number_format = FMT_EUR
            ws['F' + str(fila)].number_format = FMT_EUR
            ws['G' + str(fila)].number_format = FMT_ENT
            motor.verde(ws, 'B' + str(fila) + ':G' + str(fila))

        for b, (canal, ini) in enumerate(BLOQUES):
            m = str(FILA_MIRROR + b)          # fila del espejo de parámetros
            com, pack, obj = '$C$' + m, '$D$' + m, '$E$' + m
            c0, c1, c2 = _col(ini, 0), _col(ini, 1), _col(ini, 2)
            c3, c4, c5, c6 = (_col(ini, 3), _col(ini, 4), _col(ini, 5),
                              _col(ini, 6))
            motor.f(ws, c0 + str(fila),
                    '=IFERROR(IF(OR($C{f}="",$D{f}=""),"",'
                    '($C{f}+{pack})/($D{f}*(1-{com}))),"")'
                    .format(f=fila, pack=pack, com=com), fmt=FMT_PCT)
            motor.f(ws, c1 + str(fila),
                    '=IFERROR(IF($C{f}="","",'
                    '($C{f}+{pack})/({obj}*(1-{com}))),"")'
                    .format(f=fila, pack=pack, obj=obj, com=com), fmt=FMT_EUR)
            motor.f(ws, c2 + str(fila),
                    '=IFERROR(IF(OR(${c1}{f}="",$E{f}=""),"",${c1}{f}*(1+'
                    'INDEX(Parámetros!$B$6:$D$8,'
                    'MATCH($B${m},Parámetros!$A$6:$A$8,0),'
                    'MATCH($E{f},Parámetros!$B$5:$D$5,0)))),"")'
                    .format(c1=c1, f=fila, m=m), fmt=FMT_EUR)
            # En sala «excluir» no aplica —es tu propio comedor, no una app
            # de la que retirar el plato— así que lleva un texto propio; en
            # take away y delivery sí puedes excluirlo del canal.
            texto_no = ('No: no llega al precio que acepta el mercado'
                        if canal == 'Sala' else 'No: excluir o reformular')
            motor.f(ws, c3 + str(fila),
                    '=IF(OR(${c1}{f}="",$F{f}=""),"",IF(${c1}{f}<=$F{f},'
                    '"Sí","{no}"))'
                    .format(c1=c1, f=fila, no=texto_no))
            motor.f(ws, c4 + str(fila),
                    '=IFERROR(IF(OR(${c1}{f}="",$F{f}=""),"",'
                    'MIN(${c1}{f},$F{f})),"")'.format(c1=c1, f=fila),
                    fmt=FMT_EUR)
            motor.f(ws, c5 + str(fila),
                    '=IFERROR(IF(${c4}{f}="","",${c4}{f}*(1-{com})-{pack}),"")'
                    .format(c4=c4, f=fila, com=com, pack=pack), fmt=FMT_EUR)
            motor.f(ws, c6 + str(fila),
                    '=IFERROR(IF(OR(${c5}{f}="",$C{f}="",$G{f}=""),"",'
                    '(${c5}{f}-$C{f})*$G{f}),"")'.format(c5=c5, f=fila),
                    fmt=FMT_EUR)

    # --- fila de totales ---
    motor.val(ws, 'B' + str(FILA_TOT), 'TOTAL / MEDIA', bold=True)
    motor.f(ws, 'G' + str(FILA_TOT),
            '=IF(COUNT(G{a}:G{b})=0,"",SUM(G{a}:G{b}))'
            .format(a=FILA_INI, b=FILA_FIN), fmt=FMT_ENT, bold=True)
    for canal, ini in BLOQUES:
        c0, c6 = _col(ini, 0), _col(ini, 6)
        motor.f(ws, c0 + str(FILA_TOT),
                '=IFERROR(IF(COUNT({c}{a}:{c}{b})=0,"",'
                'AVERAGE({c}{a}:{c}{b})),"")'
                .format(c=c0, a=FILA_INI, b=FILA_FIN), fmt=FMT_PCT, bold=True)
        motor.f(ws, c6 + str(FILA_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=c6, a=FILA_INI, b=FILA_FIN), fmt=FMT_EUR, bold=True)
    fila_total(ws, FILA_TOT, 'A', _col('V', 6))

    # --- espejo de parámetros (misma hoja: el formato condicional no cruza) ---
    seccion(ws, 'B' + str(FILA_TOT + 2),
            'PARÁMETROS EN VIGOR — se leen de la hoja «Parámetros»; '
            'cámbialos allí, no aquí')
    cabecera(ws, FILA_TOT + 3, [('B', 'Canal'), ('C', 'Comisión (%)'),
                      ('D', 'Packaging por plato (€)'),
                      ('E', 'Food cost objetivo (%)')], altura=32)
    for b, canal in enumerate(CANALES):
        fila = FILA_MIRROR + b
        origen = 16 + b
        motor.f(ws, 'B' + str(fila), '=Parámetros!$A$%d' % origen, bold=True)
        motor.f(ws, 'C' + str(fila),
                '=IF(Parámetros!$B${0}="","",Parámetros!$B${0})'.format(origen),
                fmt=FMT_PCT)
        motor.f(ws, 'D' + str(fila),
                '=IF(Parámetros!$F${0}="","",Parámetros!$F${0})'.format(origen),
                fmt=FMT_EUR)
        motor.f(ws, 'E' + str(fila),
                '=IF(Parámetros!$E${0}="","",Parámetros!$E${0})'.format(origen),
                fmt=FMT_PCT)
    for fila in range(FILA_MIRROR, FILA_MIRROR + 3):
        for letra in 'BCDE':
            ws[letra + str(fila)].fill = PatternFill('solid', fgColor=GRIS)

    # --- semáforos ---
    for b, (canal, ini) in enumerate(BLOQUES):
        m = str(FILA_MIRROR + b)
        c0, c3 = _col(ini, 0), _col(ini, 3)
        rango = '{c}{a}:{c}{b}'.format(c=c0, a=FILA_INI, b=FILA_FIN)
        cf_expresion(ws, rango,
                     '=AND(ISNUMBER(${c}{a}),ISNUMBER($E${m}),${c}{a}>$E${m})'
                     .format(c=c0, a=FILA_INI, m=m),
                     motor.CF_ROJO_BG, motor.CF_ROJO_FG)
        cf_expresion(ws, rango,
                     '=AND(ISNUMBER(${c}{a}),ISNUMBER($E${m}),${c}{a}<=$E${m})'
                     .format(c=c0, a=FILA_INI, m=m),
                     motor.CF_VERDE_BG, motor.CF_VERDE_FG)
        texto_no = ('No: no llega al precio que acepta el mercado'
                    if canal == 'Sala' else 'No: excluir o reformular')
        cf_texto(ws, '{c}{a}:{c}{b}'.format(c=c3, a=FILA_INI, b=FILA_FIN),
                 (('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
                  (texto_no, motor.CF_ROJO_BG, motor.CF_ROJO_FG)))

    # --- validaciones ---
    filas = list(range(FILA_INI, FILA_FIN + 1))
    motor.dv_lista(ws, ['E%d' % r for r in filas], list(TIPOS),
                   titulo='Tipo de producto')
    motor.dv_numerica(ws, ['C%d' % r for r in filas]
                      + ['D%d' % r for r in filas]
                      + ['F%d' % r for r in filas], minimo=0,
                      titulo='Importe (€)')
    motor.dv_numerica(ws, ['G%d' % r for r in filas], minimo=0,
                      titulo='Unidades vendidas')

    ws.freeze_panes = 'H5'
    pagina(ws, titulos='$3:$4')
    return ws


# --------------------------------------------------------------------------
# Hoja Resumen
# --------------------------------------------------------------------------
def hoja_resumen(wb):
    ws = wb.create_sheet('Resumen')
    anchos(ws, {'A': 18, 'B': 14, 'C': 18, 'D': 20, 'E': 16, 'F': 16,
                'G': 18, 'H': 18, 'I': 20})
    encabezar(ws, 'Resumen por canal',
              nota='Todo se calcula desde la hoja «Carta»: aquí no se teclea '
                   'nada.')
    cabecera(ws, 4, [('A', 'Canal'),
                     ('B', 'Platos viables (n.º)'),
                     ('C', 'Platos a excluir o reformular (n.º)'),
                     ('D', 'Food cost efectivo medio al PVP de sala '
                           '(media simple, %)'),
                     ('E', 'Food cost objetivo del canal (%)'),
                     ('F', 'PVP medio necesario, sin IVA (€)'),
                     ('G', 'Margen mensual total (€)'),
                     ('H', 'Margen mensual medio por plato (€)'),
                     ('I', 'Diferencia de margen frente a la sala (€/mes)')],
             altura=60)
    for b, (canal, ini) in enumerate(BLOQUES):
        fila = 5 + b
        c0, c1, c3, c6 = (_col(ini, 0), _col(ini, 1), _col(ini, 3),
                          _col(ini, 6))
        motor.f(ws, 'A' + str(fila),
                '=Carta!$B$%d' % (FILA_MIRROR + b), bold=True)
        motor.f(ws, 'B' + str(fila),
                '=COUNTIF(Carta!${c}${a}:${c}${b},"Sí")'
                .format(c=c3, a=FILA_INI, b=FILA_FIN), fmt=FMT_ENT)
        motor.f(ws, 'C' + str(fila),
                '=IFERROR(COUNT(Carta!$C${a}:$C${b})-$B{f},"")'
                .format(a=FILA_INI, b=FILA_FIN, f=fila), fmt=FMT_ENT)
        motor.f(ws, 'D' + str(fila),
                '=IFERROR(IF(COUNT(Carta!${c}${a}:${c}${b})=0,"",'
                'AVERAGE(Carta!${c}${a}:${c}${b})),"")'
                .format(c=c0, a=FILA_INI, b=FILA_FIN), fmt=FMT_PCT)
        motor.f(ws, 'E' + str(fila),
                '=IF(Carta!$E${0}="","",Carta!$E${0})'.format(FILA_MIRROR + b),
                fmt=FMT_PCT)
        motor.f(ws, 'F' + str(fila),
                '=IFERROR(IF(COUNT(Carta!${c}${a}:${c}${b})=0,"",'
                'AVERAGE(Carta!${c}${a}:${c}${b})),"")'
                .format(c=c1, a=FILA_INI, b=FILA_FIN), fmt=FMT_EUR)
        motor.f(ws, 'G' + str(fila),
                '=IF(Carta!${c}${t}="","",Carta!${c}${t})'
                .format(c=c6, t=FILA_TOT), fmt=FMT_EUR)
        motor.f(ws, 'H' + str(fila),
                '=IFERROR(IF(OR($G{f}="",COUNT(Carta!${c}${a}:${c}${b})=0),"",'
                '$G{f}/COUNT(Carta!${c}${a}:${c}${b})),"")'
                .format(f=fila, c=c6, a=FILA_INI, b=FILA_FIN), fmt=FMT_EUR)
        if b:
            motor.f(ws, 'I' + str(fila),
                    '=IFERROR(IF(OR($G{f}="",$G$5=""),"",$G{f}-$G$5),"")'
                    .format(f=fila), fmt=FMT_EUR)
        else:
            motor.val(ws, 'I' + str(fila), 'Canal de referencia')

    cf_expresion(ws, 'D5:D7',
                 '=AND(ISNUMBER($D5),ISNUMBER($E5),$D5>$E5)',
                 motor.CF_ROJO_BG, motor.CF_ROJO_FG)
    cf_expresion(ws, 'D5:D7',
                 '=AND(ISNUMBER($D5),ISNUMBER($E5),$D5<=$E5)',
                 motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    cf_expresion(ws, 'I6:I7', '=AND(ISNUMBER($I6),$I6<0)',
                 motor.CF_ROJO_BG, motor.CF_ROJO_FG)

    seccion(ws, 'A9', 'CÓMO SE LEE')
    lecturas = [
        'Un food cost medio muy por encima del objetivo en take away o delivery '
        'no significa que el canal sea malo: significa que estás vendiendo con '
        'el precio del comedor en un canal que cobra comisión y envase.',
        'Con los datos de ejemplo, NINGÚN plato es viable en delivery: al 30 % '
        'de comisión y con un food cost objetivo del 30 %, el precio que haría '
        'falta se sale del techo en los veinte. Es el resultado normal de '
        'llevar la carta del comedor a una app sin rediseñarla, y la salida no '
        'es bajar el objetivo a ojo, sino una carta de delivery más corta, con '
        'platos de coste bajo y con el envase ya metido en el precio.',
        'En sala, un «No» no quiere decir que retires el plato: quiere decir '
        'que ese plato no llega a tu food cost objetivo al precio que acepta el '
        'mercado. Los platos de coste alto (chuletón, pescado de lonja) se '
        'ponen a precio por MARGEN, no por factor: eso lo resuelve el libro '
        '«precio-objetivo-multi-metodo.xlsx».',
        'Si un plato necesita un precio que el mercado no paga, tienes cuatro '
        'salidas: reformularlo para bajar el coste, subirlo hasta donde llegue '
        'el techo y aceptar peor food cost, sacarlo de ese canal, o sustituirlo '
        'por otro plato pensado para ese canal.',
        'La columna «Margen mensual» ya cobra el precio que puedes cobrar de '
        'verdad (el menor entre el necesario y el techo): por eso el margen de '
        'delivery puede quedarse muy por debajo del de sala aunque vendas lo '
        'mismo.',
        'El margen de esta hoja es margen de CONTRIBUCIÓN (precio neto menos '
        'coste de producto). No descuenta personal ni estructura: eso lo mide '
        'el libro «cuadro-de-mando-prime-cost.xlsx».',
    ]
    fila = 10
    for texto in lecturas:
        motor.val(ws, 'A' + str(fila), texto, wrap=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
    pagina(ws)
    return ws


# --------------------------------------------------------------------------
# Mapa de celdas
# --------------------------------------------------------------------------
def mapa():
    filas = [FILA_INI, FILA_FIN]
    m = {'fichero': NOMBRE + '.xlsx', 'hojas': {}}
    m['hojas']['Parámetros'] = {
        'celdas': {
            'IVA repercutido en sala, comida': 'B6',
            'IVA repercutido en sala, refresco azucarado': 'C6',
            'IVA repercutido en sala, bebida alcohólica': 'D6',
            'IVA repercutido en take away, comida': 'B7',
            'IVA repercutido en take away, refresco azucarado': 'C7',
            'IVA repercutido en take away, bebida alcohólica': 'D7',
            'IVA repercutido en delivery, comida': 'B8',
            'IVA repercutido en delivery, refresco azucarado': 'C8',
            'IVA repercutido en delivery, bebida alcohólica': 'D8',
            'Comisión de la plataforma en sala': 'B16',
            'Comisión de la plataforma en take away': 'B17',
            'Comisión de la plataforma en delivery': 'B18',
            'Packaging por pedido en take away': 'C17',
            'Packaging por pedido en delivery': 'C18',
            'Platos por pedido en take away': 'D17',
            'Platos por pedido en delivery': 'D18',
            'Food cost objetivo en sala': 'E16',
            'Food cost objetivo en take away': 'E17',
            'Food cost objetivo en delivery': 'E18',
            'Packaging por plato en sala': 'F16',
            'Packaging por plato en take away': 'F17',
            'Packaging por plato en delivery': 'F18',
        },
        'tablas': [
            {'titulo': 'Matriz de IVA repercutido por canal y tipo de producto',
             'cols': [['Canal', 'A', 'txt'], ['Comida', 'B', 'pct1'],
                      ['Refresco/azucarada', 'C', 'pct1'],
                      ['Bebida alcohólica', 'D', 'pct1']],
             'filas': [6, 8]},
            {'titulo': 'Parámetros del simulador por canal',
             'cols': [['Canal', 'A', 'txt'],
                      ['Comisión de la plataforma (%)', 'B', 'pct1'],
                      ['Packaging (€/pedido)', 'C', 'eur'],
                      ['Platos por pedido (n.º)', 'D', 'num1'],
                      ['Food cost objetivo (%)', 'E', 'pct1'],
                      ['Packaging por plato (€)', 'F', 'eur']],
             'filas': [16, 18]},
            {'titulo': 'Comisiones de las plataformas — orden de magnitud',
             'cols': [['Plataforma', 'A', 'txt'],
                      ['Comisión y cuotas de referencia', 'B', 'txt']],
             'filas': [24, 27]},
        ],
    }
    celdas_carta = {
        'Unidades vendidas al mes, total de la carta': 'G' + str(FILA_TOT),
    }
    for canal, ini in BLOQUES:
        c = canal.lower()
        celdas_carta['Food cost efectivo medio en %s' % c] = \
            _col(ini, 0) + str(FILA_TOT)
        celdas_carta['Margen mensual total en %s' % c] = \
            _col(ini, 6) + str(FILA_TOT)
    for i, (pid, nombre, familia, coste, pvp, uds) in enumerate(D.PLATOS):
        fila = FILA_INI + i
        celdas_carta['Coste por ración de %s' % nombre] = 'C' + str(fila)
        celdas_carta['PVP en sala sin IVA de %s' % nombre] = 'D' + str(fila)
        for canal, ini in BLOQUES:
            c = canal.lower()
            celdas_carta['Food cost efectivo en %s de %s' % (c, nombre)] = \
                _col(ini, 0) + str(fila)
            celdas_carta['PVP sin IVA necesario en %s de %s' % (c, nombre)] = \
                _col(ini, 1) + str(fila)
            celdas_carta['PVP con IVA en %s de %s' % (c, nombre)] = \
                _col(ini, 2) + str(fila)
            celdas_carta['¿Viable en %s? %s' % (c, nombre)] = \
                _col(ini, 3) + str(fila)
            celdas_carta['Margen mensual en %s de %s' % (c, nombre)] = \
                _col(ini, 6) + str(fila)
    tablas_carta = [
        {'titulo': 'Carta: datos de entrada',
         'cols': [['#', 'A', 'num'], ['Plato', 'B', 'txt'],
                  ['Coste por ración, sin IVA (€)', 'C', 'eur'],
                  ['PVP en sala, sin IVA (€)', 'D', 'eur'],
                  ['Tipo de producto (IVA)', 'E', 'txt'],
                  ['Precio techo del mercado, sin IVA (€)', 'F', 'eur'],
                  ['Uds vendidas al mes (n.º)', 'G', 'num']],
         'filas': filas},
    ]
    for canal, ini in BLOQUES:
        tablas_carta.append(
            {'titulo': 'Simulación en %s' % canal,
             'cols': [['Plato', 'B', 'txt'],
                      ['Food cost efectivo al PVP de sala (%)',
                       _col(ini, 0), 'pct1'],
                      ['PVP sin IVA necesario para el food cost objetivo (€)',
                       _col(ini, 1), 'eur'],
                      ['PVP con IVA para el cliente (€)', _col(ini, 2), 'eur'],
                      ['¿Viable? (PVP necesario vs precio techo)',
                       _col(ini, 3), 'txt'],
                      ['PVP que puedes cobrar, sin IVA (€)',
                       _col(ini, 4), 'eur'],
                      ['Ingreso neto por plato (€)', _col(ini, 5), 'eur'],
                      ['Margen mensual (€)', _col(ini, 6), 'eur']],
             'filas': filas})
    m['hojas']['Carta'] = {'celdas': celdas_carta, 'tablas': tablas_carta}
    m['hojas']['Resumen'] = {
        'celdas': {
            'Platos viables en sala': 'B5',
            'Platos viables en take away': 'B6',
            'Platos viables en delivery': 'B7',
            'Platos a excluir en sala': 'C5',
            'Platos a excluir en take away': 'C6',
            'Platos a excluir en delivery': 'C7',
            'Food cost efectivo medio en sala': 'D5',
            'Food cost efectivo medio en take away': 'D6',
            'Food cost efectivo medio en delivery': 'D7',
            'Food cost objetivo de sala': 'E5',
            'Food cost objetivo de take away': 'E6',
            'Food cost objetivo de delivery': 'E7',
            'PVP medio necesario en sala': 'F5',
            'PVP medio necesario en take away': 'F6',
            'PVP medio necesario en delivery': 'F7',
            'Margen mensual total en sala': 'G5',
            'Margen mensual total en take away': 'G6',
            'Margen mensual total en delivery': 'G7',
            'Margen mensual medio por plato en sala': 'H5',
            'Margen mensual medio por plato en take away': 'H6',
            'Margen mensual medio por plato en delivery': 'H7',
            'Diferencia de margen del take away frente a la sala': 'I6',
            'Diferencia de margen del delivery frente a la sala': 'I7',
        },
        'tablas': [
            {'titulo': 'Resumen por canal',
             'cols': [['Canal', 'A', 'txt'],
                      ['Platos viables (n.º)', 'B', 'num'],
                      ['Platos a excluir o reformular (n.º)', 'C', 'num'],
                      ['Food cost efectivo medio al PVP de sala '
                       '(media simple, %)', 'D', 'pct1'],
                      ['Food cost objetivo del canal (%)', 'E', 'pct1'],
                      ['PVP medio necesario, sin IVA (€)', 'F', 'eur'],
                      ['Margen mensual total (€)', 'G', 'eur'],
                      ['Margen mensual medio por plato (€)', 'H', 'eur'],
                      ['Diferencia de margen frente a la sala (€/mes)',
                       'I', 'eur']],
             'filas': [5, 7]},
        ],
    }
    return m


# --------------------------------------------------------------------------
def main():
    import json
    wb = Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_parametros(wb)
    hoja_carta(wb)
    hoja_resumen(wb)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = (PRODUCTO + ' · Versión 1.0 · septiembre 2026')

    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    destino = os.path.join(AQUI, 'build')
    os.makedirs(destino, exist_ok=True)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')
    wb.save(ruta)
    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)
    print('escrito:', ruta)
    print('fórmulas registradas:', len(motor.REGISTRO))


if __name__ == '__main__':
    main()
