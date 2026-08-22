#!/usr/bin/env python3
"""
inject_cache.py — Rellena el valor cacheado de las formulas en los .xlsx de
productos digitales generados con openpyxl.

PROBLEMA que resuelve: openpyxl escribe las formulas con un valor cacheado
VACIO (<f>...</f><v></v>). Excel de escritorio recalcula al abrir (con
fullCalcOnLoad="1"), pero los visores que NO recalculan (movil, previsualizador
macOS/Quick Look, Google Sheets a veces, Excel en modo manual) muestran las
celdas de resultado EN BLANCO -> el cliente percibe "valores fijos, no se
actualiza". Incidente real 2026-07-21 (Lucas Amorim, Plan de Negocio
Bar-Restaurante). Ver memoria feedback-openpyxl-cache-valores-excel.

QUE HACE: evalua cada formula con pycel e inyecta el valor como <v> en el XML,
preservando la formula (<f>) y el formato -> el fichero abre mostrando valores
Y sigue recalculando al editar.

USO:  python3 scripts/inject_cache.py fichero1.xlsx fichero2.xlsx ...
DEP:  pip install pycel   (arrastra numpy, networkx; instalacion ligera)

LIMITES conocidos (funciones que pycel NO implementa) -> sustituir la formula
por un equivalente editable ANTES de ejecutar:
  - PMT(rate,nper,pv)  ->  =pv*(rate/12)/(1-(1+rate/12)^(-nper*12))   (sistema frances)
  - COUNTA(rango)      ->  SUM(rango_de_recuentos)  o cachear el conteo a mano
Las celdas de formula que pycel no pueda evaluar quedan sin cache (fallos_pycel>0).

VERIFICAR SIEMPRE despues: abrir con openpyxl.load_workbook(f, data_only=True) y
confirmar que las celdas de resultado tienen numero (no None). Contar acentos con
openpyxl, NO con regex sobre el XML (openpyxl escribe tildes como entidades &#243;).
"""
# 2026-08-18 (AICP): openpyxl 3.1.x escribe el cache vacío como <v /> (autocerrado) y no
# como <v></v> → las regex aceptan ambas formas. Antes inyectaba 0 celdas en silencio.
# 2026-08-22 (AICP): pycel devuelve a veces escalares de numpy (np.float64,
# np.str_) y mas abajo solo se inyectan int/float/str nativos -> la celda se
# quedaba sin cache en silencio. _nativo() los convierte antes de clasificar.
import openpyxl, zipfile, re, shutil, os, html, logging, contextlib
logging.disable(logging.CRITICAL)
from pycel import ExcelCompiler

try:
    import numpy as _np
except ImportError:
    _np = None


def _nativo(v):
    """Escalar de numpy -> tipo nativo de Python (idempotente con el resto)."""
    if _np is not None:
        if isinstance(v, _np.generic):
            return v.item()
        if isinstance(v, _np.ndarray) and v.ndim == 0:
            return v.item()
    return v

def inject(path):
    wb = openpyxl.load_workbook(path)
    # recopilar celdas de formula por hoja
    formula_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                # 2026-08-22 (AICP): por TIPO de celda, no por el '=' inicial: las etiquetas
                # «= Margen bruto» convertidas a texto por postprocess-transversal.py no son formulas.
                if c.data_type == 'f' and isinstance(c.value, str) and c.value.startswith('='):
                    formula_cells.append((ws.title, c.coordinate))
    # evaluar con pycel
    xl = ExcelCompiler(filename=path)
    vals = {}   # (sheet, ref) -> value
    failed = 0
    for sheet, ref in formula_cells:
        try:
            with open(os.devnull,'w') as dn, contextlib.redirect_stderr(dn):
                v = _nativo(xl.evaluate(f"'{sheet}'!{ref}"))
            if v is not None and not (isinstance(v, str) and v == ''):
                vals[(sheet, ref)] = v
        except Exception:
            failed += 1
    # mapear nombre de hoja -> worksheets/sheetN.xml
    z = zipfile.ZipFile(path)
    wbxml = z.read('xl/workbook.xml').decode('utf-8')
    rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    relmap = {}
    for rel in re.findall(r'<Relationship[^>]*/>', rels):
        rid = re.search(r'Id="([^"]+)"', rel); tgt = re.search(r'Target="([^"]+)"', rel)
        if rid and tgt: relmap[rid.group(1)] = tgt.group(1)
    sheet_to_file = {}
    for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wbxml):
        name, rid = m.group(1), m.group(2)
        tgt = relmap.get(rid, '')
        if tgt:
            sfile = tgt.lstrip('/') if tgt.startswith('/') else 'xl/' + tgt
            sheet_to_file[html.unescape(name)] = sfile
    # editar cada sheet xml
    parts = {n: z.read(n) for n in z.namelist()}
    injected = 0
    for sheet, sfile in sheet_to_file.items():
        if sfile not in parts: continue
        xml = parts[sfile].decode('utf-8')
        cellvals = {ref: v for (s, ref), v in vals.items() if s == sheet}
        for ref, v in cellvals.items():
            if isinstance(v, bool): continue
            if isinstance(v, (int, float)):
                pat = re.compile(r'(<c r="'+ref+r'"[^>]*>)(<f[^>]*>[^<]*</f>)(?:<v>[^<]*</v>|<v\s*/>)?(</c>)')
                rep = r'\1\2<v>'+repr(float(v))+r'</v>\3'
                xml, n = pat.subn(rep, xml, count=1)
                injected += n
            elif isinstance(v, str):
                esc = html.escape(v)
                pat = re.compile(r'<c r="'+ref+r'"([^>]*)>(<f[^>]*>[^<]*</f>)(?:<v>[^<]*</v>|<v\s*/>)?</c>')
                # IDEMPOTENTE (2026-08-22): si la celda ya lleva t="str" de una pasada anterior, se
                # quita antes de reinsertarlo; si no, la segunda pasada dejaba t="str" t="str"
                # (atributo duplicado → XML inválido → openpyxl revienta). Cazado en el 10 y el 12.
                def rep(m, esc=esc, ref=ref):
                    attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
                    return '<c r="'+ref+'"'+attrs+' t="str">'+m.group(2)+'<v>'+esc+'</v></c>'
                xml, n = pat.subn(rep, xml, count=1)
                injected += n
        parts[sfile] = xml.encode('utf-8')
    # asegurar fullCalcOnLoad
    if 'xl/workbook.xml' in parts:
        w = parts['xl/workbook.xml'].decode('utf-8')
        if '<calcPr' in w and 'fullCalcOnLoad' not in w:
            w = re.sub(r'<calcPr([^>]*)/>', r'<calcPr\1 fullCalcOnLoad="1"/>', w)
        elif '<calcPr' not in w:
            w = w.replace('</workbook>', '<calcPr calcId="124519" fullCalcOnLoad="1"/></workbook>')
        parts['xl/workbook.xml'] = w.encode('utf-8')
    z.close()
    tmp = path + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n, data in parts.items():
            zout.writestr(n, data)
    shutil.move(tmp, path)
    return len(formula_cells), injected, failed

if __name__ == '__main__':
    import sys
    for p in sys.argv[1:]:
        nf, inj, fail = inject(p)
        print(f'  {os.path.basename(p)}: formulas={nf} cache_inyectado={inj} fallos_pycel={fail}')
