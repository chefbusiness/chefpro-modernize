#!/usr/bin/env python3
"""
grupo_a.py — §2 de `pack-appcc-v2-SPEC.md`: los registros con MEDICIÓN
(01, 02, 05, 06, 09, 10). Son los seis ficheros del pack en los que una
fórmula emite un veredicto, así que son los que pueden dar por bueno un dato
fuera de límite — el fallo que la ronda 1 marcó como «no listo».

  · 01 temperaturas diario — congeladores `<=-18` SIN suelo (DOM-09/TEC-05/
    COM-18) y caliente `>=65` SIN techo (TEC-31); rótulos coherentes con las
    Instrucciones; columnas Hora M / Firma M / Hora T / Firma T y «Nº
    incidencia (→ 11)» (DOM-34); semáforo y resumen semanal de desviaciones.
  · 02 recepción temperaturas — hoja auxiliar «Límites» con los ocho límites
    legales por familia, columna «Familia de producto» con desplegable que
    valida y «Límite máx.» resuelto por VLOOKUP (DOM-04/TEC-01, altas); aviso
    rojo si Estado = RECHAZAR y Aceptado = S (TEC-30).
  · 05 recepción mercancías — «Nº albarán», «Nº lote» y «Firma receptor»
    (DOM-23) y la nota que explica por qué NO se fusiona con el 02.
  · 06 trazabilidad — pestaña de salida / uso interno (DOM-22/TEC-24/COM-21) y
    el «< 4 horas» inventado sustituido por lo que dice el Reg. 178/2002.
  · 09 aceite — `>=25` CAMBIAR y `>=20` VIGILAR (TEC-06), la temperatura DENTRO
    del veredicto (TEC-07/DOM-25/COM-19) y pestaña de retirada por gestor.
  · 10 agua — guardas invertidas (INCOMPLETO / FALTA CLORO / REVISAR)
    (DOM-32/TEC-15), 31 filas y la frecuencia en Instrucciones (COM-24).

Se ejecuta en dos tiempos alrededor del motor, igual que en
`kit-escandallos-v2_0/grupo_a.py`:
  pre(wb, fname, cambios)   → inserciones estructurales (columnas nuevas),
                              detrás de un centinela para que la 2.ª pasada no
                              vuelva a insertar.
  post(wb, fname, cambios, registro) → la rejilla entera: cabeceras, fórmulas,
                              replicado, DV, semáforo, ejemplos y pies.
"""
import contextlib
import os
import shutil

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break, RowBreak

import motor

FICHEROS = [
    '01-registro-temperaturas-diario.xlsx',
    '02-registro-temperaturas-recepcion.xlsx',
    '05-checklist-recepcion-mercancias.xlsx',
    '06-registro-trazabilidad.xlsx',
    '09-control-aceite-fritura.xlsx',
    '10-control-agua-potable.xlsx',
]

DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado',
        'Domingo']

# ==========================================================================
# 01 — Registro de temperaturas diario
# ==========================================================================
# (fila de la banda, rótulo, tipo de límite). El rótulo es el MISMO texto que
# la hoja de Instrucciones: COM-18 nació justo de que no lo fueran (las
# Instrucciones decían «-18 °C o inferior» y la fórmula exigía un rango
# cerrado -25/-18, así que un arcón a -26 °C —conforme— salía ALERTA y el pie
# de la hoja mandaba abrir una acción correctiva por cada falsa desviación).
BLOQUES_01 = [
    (5, 'Cámara 1 (Refrigeración)', 'Rango: 0 °C a 4 °C', 'refri4'),
    (15, 'Cámara 2 (Refrigeración)', 'Rango: 0 °C a 4 °C', 'refri4'),
    (25, 'Congelador 1', 'Rango: -18 °C o inferior', 'cong'),
    (35, 'Congelador 2', 'Rango: -18 °C o inferior', 'cong'),
    (45, 'Exposición Fría (Vitrina)', 'Rango: 0 °C a 8 °C', 'refri8'),
    (55, 'Exposición Caliente (Baño María)', 'Rango: 65 °C o superior',
     'caliente'),
]

# DOM-R2-26 (ronda 2): la v2.0 desdobló la lectura de la mañana y la de la
# tarde (Hora M / Firma M / Hora T / Firma T) pero dejó UNA sola columna de
# «Nº incidencia» por fila, y cada fila es un día con DOS lecturas que pueden
# desviarse las dos —la cámara a 6,5 °C por la mañana y a 7 por la tarde son
# dos incidencias distintas en el registro 11—. El contador C65 ya las cuenta
# por separado; ahora también se pueden referenciar por separado.
CAB_01 = ['Día', 'Temp. Mañana (°C)', 'Estado', 'Hora M', 'Firma M',
          'Nº inc. M (→ 11)', 'Temp. Tarde (°C)', 'Estado', 'Hora T',
          'Firma T', 'Nº inc. T (→ 11)']
ANCHOS_01 = [30, 14, 12, 9, 14, 16, 14, 12, 9, 14, 16]
NC_01 = len(CAB_01)


def _formula_01(tipo, col_temp, fila):
    ref = f'${col_temp}{fila}'
    if tipo == 'refri4':
        return f'=IF({ref}="","",IF(AND({ref}>=0,{ref}<=4),"OK","ALERTA"))'
    if tipo == 'refri8':
        return f'=IF({ref}="","",IF(AND({ref}>=0,{ref}<=8),"OK","ALERTA"))'
    if tipo == 'cong':
        # DOM-09/TEC-05/COM-18: sin suelo. El requisito legal es «-18 °C o
        # inferior»; cuanto más frío, mejor.
        return f'=IF({ref}="","",IF({ref}<=-18,"OK","ALERTA"))'
    # TEC-31: sin techo. Un baño maría a 102 °C no es una desviación.
    return f'=IF({ref}="","",IF({ref}>=65,"OK","ALERTA"))'


def _post_01(wb, fname, cambios):
    ws = wb['Registro Semanal']
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for fila in range(64, max(71, ws.max_row + 1)):
        for col in range(1, NC_01 + 2):
            ws.cell(row=fila, column=col).value = None
    ws.row_breaks = RowBreak()

    ws['A1'] = 'Registro de Temperaturas — Control Diario'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells(f'A1:{get_column_letter(NC_01)}1')
    ws['A2'] = 'Semana del: ___/___/______ al ___/___/______'
    ws['A3'] = ('Responsable: _________________________________     '
                'Termómetro / sonda nº: ______________')

    rangos_alerta, rangos_temp = [], []
    for n_bloque, (banda_fila, rotulo, limite, tipo) in enumerate(BLOQUES_01):
        hdr = banda_fila + 1
        d0, d1 = hdr + 1, hdr + 7
        motor.banda(ws, banda_fila, f'📍 {rotulo}  |  {limite}', NC_01)
        motor.cabecera(ws, hdr, CAB_01, ANCHOS_01)
        # TEC-14: el 01 es la única hoja de registro cuyas filas de título de
        # impresión (1:3) NO llevan encabezado de columna —lo lleva cada bloque
        # dentro—, así que un bloque partido por el salto de página deja filas
        # firmadas sin cabecera. Se fuerza el salto ANTES de cada bloque a
        # partir del segundo: así ninguno se parte nunca.
        if n_bloque:
            ws.row_breaks.append(Break(id=banda_fila - 1))

        # fila plantilla + replicado a los 7 días (§1.3)
        for col in range(1, NC_01 + 1):
            cel = ws.cell(row=d0, column=col)
            cel.value = None
            if col in (2, 7):
                motor.verde(cel, motor.FMT_TEMP)
            elif col in (3, 8):
                motor.calculada(cel)
            elif col == 1:
                motor.calculada(cel)
                cel.alignment = Alignment(horizontal='left', vertical='center')
            else:
                motor.verde(cel, '@')
        f_man = _formula_01(tipo, 'B', d0)
        f_tar = _formula_01(tipo, 'G', d0)
        ws.cell(row=d0, column=3).value = f_man
        ws.cell(row=d0, column=8).value = f_tar
        motor.reg(ws, f'C{d0}', f_man)
        motor.reg(ws, f'H{d0}', f_tar)
        motor.replicar_filas(ws, d0, d0, d1, ncols=NC_01, alto=18)
        for i, dia in enumerate(DIAS):
            ws.cell(row=d0 + i, column=1).value = dia

        motor.semaforo(ws, motor.rango('C', d0, d1))
        motor.semaforo(ws, motor.rango('H', d0, d1))
        rangos_alerta += [motor.rango('C', d0, d1), motor.rango('H', d0, d1)]
        rangos_temp += [motor.rango('B', d0, d1), motor.rango('G', d0, d1)]

    # §1.2 — el 01 no tiene desplegables (las temperaturas son numéricas), pero
    # sí puede tener un dedo torpe: un «35» en un congelador o un «6,5» tecleado
    # como «65» pasaban sin más. La validación acota a un rango físicamente
    # posible en una cocina y deja el resto al veredicto de la columna Estado.
    motor.dv_decimal(ws, rangos_temp, 'Temperatura leída',
                     'Introduce la temperatura en °C, entre -40 y 130. Si el '
                     'valor es correcto y aun así se rechaza, revisa que no '
                     'estés escribiendo el punto decimal como separador de '
                     'miles.', minimo=-40, maximo=130)

    # §1.5 — ejemplos sembrados en el primer bloque (Cámara 1)
    motor.sembrar(ws, 7, {'B': 3.2, 'D': '08:15', 'E': 'A.R.', 'G': 3.8,
                          'I': '19:40', 'J': 'M.S.'}, marca_col='K')
    motor.sembrar(ws, 8, {'B': 2.8, 'D': '08:10', 'E': 'A.R.', 'G': 4.0,
                          'I': '19:35', 'J': 'M.S.'}, marca_col='K')
    motor.sembrar(ws, 9, {'B': 6.5, 'D': '08:20', 'E': 'A.R.',
                          'F': 'INC-001', 'G': 3.5,
                          'I': '19:45', 'J': 'M.S.'}, marca_col='K')
    cambios.append('01: 3 filas de ejemplo en Cámara 1, una con desviación '
                   'encadenada al registro 11 (DOM-03/COM-09); columnas de '
                   'incidencia desdobladas en mañana y tarde (DOM-R2-26)')

    # §1.3 — resumen de la semana con COUNTIF sobre TODOS los rangos
    ws.cell(row=65, column=1,
            value='Lecturas fuera de límite en la semana (todas las secciones)'
            ).font = Font(bold=True, size=11)
    suma = '+'.join(f'COUNTIF({r},"ALERTA")' for r in rangos_alerta)
    f = f'={suma}'
    ws.cell(row=65, column=3).value = f
    motor.calculada(ws.cell(row=65, column=3), motor.FMT_ENT)
    ws.cell(row=65, column=3).font = Font(bold=True, size=12)
    motor.reg(ws, 'C65', f)
    motor.cf_formula(ws, 'C65', '=C65>0')

    motor.nota(ws, 66, 'Si aparece ALERTA: registrar la incidencia en el '
                       'registro 11 (Acciones Correctivas) con su nº, anotarlo '
                       'en «Nº inc. M» o «Nº inc. T» según la lectura que se '
                       'haya desviado, y avisar al responsable.',
               ncols=NC_01)
    motor.nota(ws, 67, motor.CONSERVACION, ncols=NC_01)
    motor.nota(ws, 68, motor.MARCA, ncols=NC_01)
    # El 01 tiene SEIS bloques con su propia cabecera cada diez filas: repetir
    # la del primero pondría «Cámara 1» encima de las filas del congelador en
    # la página 2. Lo que se repite es el encabezado del documento (1:3), y el
    # salto de página forzado (TEC-14) impide que un bloque se parta.
    motor.IMPRESION[(fname, ws.title)] = ('1:3', True, False)

    motor.escribir_instrucciones(wb, 'Registro de Temperaturas — Control Diario', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Registra la temperatura de cada equipo 2 veces al día '
              '(mañana y tarde) en las celdas verdes.'),
        ('b', 'Cada lectura tiene su propia hora y su propia firma: '
              'Hora M / Firma M para la mañana y Hora T / Firma T para la '
              'tarde. Si los turnos son distintos, la hoja acredita quién '
              'tomó cada medida.'),
        ('b', 'Las columnas «Estado» se calculan solas y se pintan: verde OK, '
              'rojo ALERTA. No se escriben a mano.'),
        ('b', 'Cada lectura tiene su propia columna de incidencia: «Nº inc. M» '
              'para la mañana y «Nº inc. T» para la tarde. Las dos enlazan con '
              'el registro 11 de Acciones Correctivas — un mismo día puede '
              'tener dos desviaciones distintas, y cada una es una incidencia.'),
        ('b', 'La celda C65 cuenta todas las lecturas fuera de límite de la '
              'semana y se pone en rojo si hay alguna.'),
        ('b', 'Imprime la hoja semanal en A4 apaisado y archívala firmada. '
              'Cada equipo tiene su propia fila de cabecera cada pocas filas, '
              'así que lo que se repite arriba de cada página es el '
              'encabezado del documento (título, semana y responsable).'),
        ('h', 'Límites de temperatura que aplica la hoja'),
        ('b', 'Cámaras de refrigeración: 0 °C a 4 °C.'),
        ('b', 'Congeladores: -18 °C o inferior. NO hay límite por abajo: un '
              'arcón a -26 °C o un abatidor a -30 °C están conformes.'),
        ('b', 'Exposición fría (vitrina): 0 °C a 8 °C.'),
        ('b', 'Exposición caliente (baño maría): 65 °C o superior. NO hay '
              'límite por arriba: 102 °C en un baño maría con agua hirviendo '
              'es normal y no es una desviación.'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Frecuencia recomendada: 2 lecturas diarias por equipo, al '
              'abrir y antes de cerrar; una tercera tras cada carga grande.'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('01: 6 bloques con Hora/Firma por turno, límites sin suelo '
                   'ni techo y semáforo (DOM-09/TEC-05/TEC-31/DOM-34/TEC-02)')


# ==========================================================================
# 02 — Recepción de temperaturas
# ==========================================================================
# DOM-04 (alta): el desplegable de v1.1 ofrecía un único «4» para «carne,
# pescado, lácteos». Con eso el registro daba por CONFORME pescado fresco a
# 4 °C (debe recibirse en hielo fundente, 0-2 °C) y carne picada a 4 °C
# (límite legal 2 °C), y RECHAZABA canales de vacuno a 5-7 °C que sí cumplen.
# El fichero dejaba por escrito la aceptación de producto fuera de límite.
LIMITES_02 = [
    ('Pescado fresco (en hielo fundente)', 2,
     'Reg. (CE) 853/2004, Anexo III, Secc. VIII: temperatura del hielo '
     'fundente (0-2 °C)'),
    # DOM-R2-06 (alta, ronda 2): el Reg. (CE) 853/2004, Anexo III, Secc. V,
    # Cap. III separa carne PICADA (≤2 °C) de PREPARADOS de carne (≤4 °C:
    # hamburguesas, salchichas frescas, brochetas, adobados — el grueso de lo
    # que recibe un restaurante). Metidos en una sola familia a 2 °C, el
    # registro escupía RECHAZAR sobre entregas legales y, si el cliente las
    # aceptaba, la regla «Estado=RECHAZAR y Aceptado=S» le pintaba la fila en
    # rojo como no conformidad asumida a sabiendas.
    ('Carne picada', 2,
     'Reg. (CE) 853/2004, Anexo III, Secc. V, Cap. III'),
    ('Preparados de carne (hamburguesas, salchichas frescas, adobados)', 4,
     'Reg. (CE) 853/2004, Anexo III, Secc. V, Cap. III'),
    ('Despojos', 3, 'Reg. (CE) 853/2004, Anexo III, Secc. I'),
    # DOM-R2-22: la caza mayor silvestre va a 7 °C (Secc. IV), no a los 4 °C de
    # las aves (Secc. II). Un corzo recibido a 6 °C es legal.
    ('Aves y caza menor', 4, 'Reg. (CE) 853/2004, Anexo III, Secc. II y IV'),
    ('Caza mayor silvestre', 7, 'Reg. (CE) 853/2004, Anexo III, Secc. IV'),
    # DOM-R2-02 / COM-R2-01: el RD 3484/2000 está derogado por el RD 1021/2022,
    # que además ya no fija temperaturas: las fija el APPCC del operador.
    ('Lácteos y platos preparados refrigerados', 4,
     'Criterio del establecimiento (RD 1021/2022: la temperatura la fija tu '
     'APPCC) y ficha del fabricante'),
    ('Canales y despieces de vacuno, ovino y porcino', 7,
     'Reg. (CE) 853/2004, Anexo III, Secc. I'),
    ('Congelados', -18, 'Reg. (CE) 853/2004 y RD 1109/1991'),
    ('Ambiente (sin control de temperatura)', 'N/A',
     'Verificar integridad del envase, etiquetado y fecha de caducidad'),
]
def lim_02(prefijo):
    """Nombre EXACTO de una familia por su prefijo.

    Las filas sembradas y los casos §6 la referenciaban por índice
    (`LIMITES_02[5][0]`); al partir «Carne picada y preparados» y «Aves y caza»
    en dos filas cada una (DOM-R2-06 / DOM-R2-22) los índices se desplazaron dos
    posiciones y la siembra habría escrito «Caza mayor silvestre» donde ponía
    «Canales». Por nombre no se puede desplazar: si el prefijo deja de existir,
    esto revienta en la generación en vez de en el entregable.
    """
    for nombre, _lim, _base in LIMITES_02:
        if nombre.startswith(prefijo):
            return nombre
    raise SystemExit(f'02: no hay familia que empiece por {prefijo!r}')


LIM_F0 = 5
LIM_LIBRES = 6
# El rango del VLOOKUP y el del desplegable llegan MÁS ABAJO que la última
# familia precargada: si acabaran en la fila 12 —la última llena—, la nota que
# invita a «añadir familias propias» sería mentira, porque no habría dónde. Con
# seis filas verdes libres, lo que el cliente escriba en A13:B18 entra solo en
# el desplegable y en la búsqueda del límite.
LIM_F1 = 4 + len(LIMITES_02) + LIM_LIBRES        # filas 5..18 de «Límites»
LIM_ULT = 4 + len(LIMITES_02)                    # última familia precargada

CAB_02 = ['Fecha', 'Producto', 'Proveedor', 'Temp. recibida (°C)',
          'Familia de producto', 'Límite máx. (°C)', 'Estado',
          'Lote / Caducidad', 'Aceptado (S/N)', 'Observaciones']
ANCHOS_02 = [14, 24, 20, 15, 38, 14, 14, 18, 13, 30]
F0_02, F1_02 = 5, 44                              # 40 filas (§1.3, mensual)


def _pre_02(wb, cambios):
    ws = wb['Recepción Temperaturas']
    if ws['E4'].value == 'Familia de producto':
        return                                    # centinela: ya está en v2
    motor.insertar_columna(ws, 5)
    cambios.append('02: columna «Familia de producto» insertada en E '
                   '(el límite pasa a calcularse, no a teclearse)')


def _hoja_limites(wb, cambios):
    ws = motor.hoja(wb, 'Límites')
    ws['A1'] = 'Límites legales de temperatura en recepción'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = ('Esta hoja alimenta el desplegable «Familia de producto» del '
                'registro. Puedes añadir familias propias en las filas verdes '
                f'libres (A{LIM_ULT + 1}:C{LIM_F1}): el desplegable y la '
                'búsqueda del límite las recogen solas.')
    ws['A3'].font = Font(size=10, italic=True)
    motor.cabecera(ws, 4, ['Familia de producto', 'Límite máx. (°C)',
                           'Base legal / nota'], [40, 16, 78])
    for r in range(LIM_F0, LIM_F1 + 1):
        i = r - LIM_F0
        familia, limite, base = LIMITES_02[i] if i < len(LIMITES_02) \
            else (None, None, None)
        ws.cell(row=r, column=1, value=familia)
        ws.cell(row=r, column=2, value=limite)
        ws.cell(row=r, column=3, value=base)
        motor.verde(ws.cell(row=r, column=1), align='left')
        # «N/A» es texto: con el formato numérico '0' Excel lo pinta igual pero
        # la celda queda mintiendo sobre su tipo. Sólo se formatea lo numérico.
        motor.verde(ws.cell(row=r, column=2),
                    motor.FMT_ENT if isinstance(limite, (int, float)) else None)
        motor.verde(ws.cell(row=r, column=3), align='left')
        ws.row_dimensions[r].height = 30
    motor.nota(ws, LIM_F1 + 2,
               '«N/A» significa que la familia no tiene límite de temperatura '
               'en recepción: el Estado del registro devuelve OK y lo que hay '
               'que verificar es el envase, el etiquetado y la caducidad.',
               ncols=3)
    motor.nota(ws, LIM_F1 + 3,
               f'Las filas A{LIM_ULT + 1}:C{LIM_F1} están libres para tus '
               'propias familias (por ejemplo, un límite más estricto pactado '
               'con un proveedor). El desplegable del registro y la búsqueda '
               'del límite ya las incluyen.', ncols=3)
    motor.nota(ws, LIM_F1 + 4, motor.MARCA, ncols=3)
    motor.IMPRESION[('02-registro-temperaturas-recepcion.xlsx', 'Límites')] = \
        (4, True)
    cambios.append(f'02: hoja «Límites» con {len(LIMITES_02)} familias y su '
                   'base legal (DOM-04)')
    return ws


def _post_02(wb, fname, cambios):
    _hoja_limites(wb, cambios)
    ws = wb['Recepción Temperaturas']
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for fila in range(F1_02 + 1, max(F1_02 + 8, ws.max_row + 1)):
        for col in range(1, 12):
            ws.cell(row=fila, column=col).value = None

    ws['A1'] = 'Control de Temperaturas — Recepción de Mercancías'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:J1')
    ws['A2'] = ('Mes: _______________    '
                'Responsable: _________________________________')
    motor.cabecera(ws, 4, CAB_02, ANCHOS_02)

    for col in range(1, 11):
        cel = ws.cell(row=F0_02, column=col)
        cel.value = None
        if col in (6, 7):
            motor.calculada(cel, motor.FMT_TEMP if col == 6 else None)
        elif col == 4:
            motor.verde(cel, motor.FMT_TEMP)
        elif col in (2, 3, 5, 10):
            motor.verde(cel, align='left')
        else:
            motor.verde(cel, '@')

    # El límite se BUSCA por familia; no se teclea. `IFERROR` deja "" si la
    # familia está vacía o no está en la tabla, y ese "" es lo que arma la
    # guarda de TEC-01 en la columna Estado.
    # Doble guarda: familia vacía → ""; familia escrita pero con el límite de
    # la hoja «Límites» sin rellenar (puede pasar en las filas libres) → ""
    # también, NUNCA 0, que es lo que devuelve VLOOKUP sobre una celda vacía y
    # convertiría cualquier refrigerado en «RECHAZAR» sin motivo.
    _vl = (f"VLOOKUP($E{F0_02},'Límites'!$A${LIM_F0}:$B${LIM_F1},2,FALSE)")
    f_lim = f'=IF($E{F0_02}="","",IFERROR(IF({_vl}="","",{_vl}),""))'
    # TEC-01 (alta): la fórmula v1.1 sólo guardaba contra D vacío. Con el
    # límite sin elegir, Excel coacciona la celda vacía a 0 y CUALQUIER
    # temperatura negativa devolvía «OK»: un congelado recibido a -5 °C —14
    # grados por encima de su límite— se registraba como conforme.
    f_est = (f'=IF(OR($D{F0_02}="",$F{F0_02}=""),"",'
             f'IF($F{F0_02}="N/A","OK",'
             f'IF($D{F0_02}<=$F{F0_02},"OK","RECHAZAR")))')
    ws.cell(row=F0_02, column=6).value = f_lim
    ws.cell(row=F0_02, column=7).value = f_est
    motor.reg(ws, f'F{F0_02}', f_lim)
    motor.reg(ws, f'G{F0_02}', f_est)
    motor.replicar_filas(ws, F0_02, F0_02, F1_02, ncols=10, alto=20)

    motor.dv_lista(
        ws, f"'Límites'!$A${LIM_F0}:$A${LIM_F1}",
        [motor.rango('E', F0_02, F1_02)],
        'Familia de producto',
        'Elige una familia de la lista. El límite legal se calcula solo a '
        'partir de ella; si escribes texto libre, la columna Estado se queda '
        'en blanco y la fila no acredita nada.',
        # TEC-15 (ronda 2): `allow_blank=False`. Excel sólo valida al escribir
        # o pegar, así que las 40 filas que se entregan vacías siguen vacías sin
        # una sola advertencia; lo que ya no se puede es dejar la celda a medias
        # tecleando algo que no esté en la lista.
        allow_blank=False)
    motor.dv_lista(ws, '"S,N"', [motor.rango('I', F0_02, F1_02)],
                   'Aceptado', 'Escribe S (aceptado) o N (rechazado).')
    motor.dv_decimal(ws, [motor.rango('D', F0_02, F1_02)],
                     'Temperatura recibida',
                     'Introduce la temperatura en °C, entre -40 y 60.',
                     minimo=-40, maximo=60)

    motor.semaforo(ws, motor.rango('G', F0_02, F1_02))
    # TEC-30: la fila entera en rojo cuando se acepta lo que la fórmula
    # rechaza. Es la combinación que un inspector busca al hojear el registro.
    motor.cf_formula(ws, f'A{F0_02}:J{F1_02}',
                     f'=AND($G{F0_02}="RECHAZAR",$I{F0_02}="S")')

    motor.sembrar(ws, 5, {'A': '05/09/2026', 'B': 'Merluza fresca',
                          'C': 'Pescados de la Ría S.L.', 'D': 1.5,
                          'E': lim_02('Pescado fresco'), 'H': 'L-2026-0912 / 07-09-26',
                          'I': 'S'}, marca_col='J')
    motor.sembrar(ws, 6, {'A': '05/09/2026', 'B': 'Canal de ternera',
                          'C': 'Cárnicas del Norte', 'D': 5.5,
                          'E': lim_02('Canales'), 'H': 'L-2026-0913 / 12-09-26',
                          'I': 'S'}, marca_col='J')
    motor.sembrar(ws, 7, {'A': '05/09/2026', 'B': 'Guisantes congelados',
                          'C': 'Congelados del Sur', 'D': -14,
                          'E': lim_02('Congelados'), 'H': 'L-2026-0914 / 03-2028',
                          'I': 'N',
                          'J': 'Devuelto al proveedor · incidencia INC-002'},
                  marca_col='J')

    motor.nota(ws, F1_02 + 2,
               'Si Estado = RECHAZAR: devolver el producto al proveedor y '
               'abrir la incidencia en el registro 11 (Acciones Correctivas). '
               'Si aun así se acepta, la fila entera se pinta en rojo: es una '
               'no conformidad aceptada a sabiendas y hay que justificarla.',
               ncols=10)
    motor.nota(ws, F1_02 + 3, motor.CONSERVACION, ncols=10)
    motor.nota(ws, F1_02 + 4, motor.MARCA, ncols=10)
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    filas_tabla = [('b', f'{fam} → máx. {lim} °C' if lim != 'N/A'
                    else f'{fam} → sin límite de temperatura (N/A)')
                   for fam, lim, _ in LIMITES_02]
    motor.escribir_instrucciones(
        wb, 'Control de Temperaturas en Recepción de Mercancías', [
            ('h', 'Cómo usar esta plantilla'),
            ('b', 'Mide la temperatura del producto al recibirlo, con sonda '
                  'entre dos envases (nunca pinchando el producto que se va a '
                  'servir) y anótala en «Temp. recibida (°C)».'),
            ('b', 'Elige la FAMILIA en el desplegable. El «Límite máx. (°C)» '
                  'no se escribe: se calcula solo desde la hoja «Límites».'),
            ('b', 'La columna «Estado» compara los dos valores y responde OK o '
                  'RECHAZAR, con semáforo. Si falta la temperatura o la '
                  'familia, se queda en blanco: nunca da OK por defecto.'),
            ('b', 'Si marcas «Aceptado = S» en una fila con Estado = RECHAZAR, '
                  'la fila entera se pinta en rojo. Es deliberado: esa '
                  'combinación hay que justificarla por escrito.'),
            ('h', 'Límites de temperatura en recepción (hoja «Límites»)'),
        ] + filas_tabla + [
            ('p', 'Puedes añadir familias propias en la hoja «Límites», dentro '
                  f'del rango A{LIM_ULT + 1}:C{LIM_F1}, que se entregan '
                  'vacías a propósito.'),
            ('h', 'Frecuencia y archivo'),
            ('b', 'Frecuencia: TODAS las entregas de producto con temperatura '
                  'regulada. La hoja trae 40 filas, suficientes para un mes de '
                  'una o dos entregas diarias.'),
            ('b', 'Este registro es el PCC de temperatura en recepción. La '
                  'verificación documental (albarán, etiquetado, envase) va en '
                  'el registro 05: son complementarios, no duplicados.'),
            ('b', motor.CONSERVACION),
        ], cambios)
    cambios.append('02: límite por familia con VLOOKUP, guarda doble en el '
                   'Estado y aviso de RECHAZAR+aceptado (DOM-04/TEC-01/TEC-30)')


# ==========================================================================
# 05 — Checklist de recepción de mercancías
# ==========================================================================
CAB_05 = ['Fecha', 'Producto', 'Proveedor', 'Temp. (°C)', 'Caducidad',
          'Etiquetado OK', 'Envase OK', 'Aspecto OK', 'Cantidad OK',
          'Aceptado', 'Nº albarán', 'Nº lote', 'Firma receptor',
          'Observaciones']
ANCHOS_05 = [14, 22, 20, 12, 14, 12, 12, 12, 12, 12, 16, 18, 18, 28]
F0_05, F1_05 = 5, 44


def _pre_05(wb, cambios):
    ws = wb['Recepción Mercancías']
    if ws['K4'].value == 'Nº albarán':
        return
    for _ in range(3):
        motor.insertar_columna(ws, 11)            # K, L, M antes de Observaciones
    cambios.append('05: columnas «Nº albarán», «Nº lote» y «Firma receptor» '
                   'insertadas antes de Observaciones (DOM-23)')


def _post_05(wb, fname, cambios):
    ws = wb['Recepción Mercancías']
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for fila in range(F1_05 + 1, max(F1_05 + 8, ws.max_row + 1)):
        for col in range(1, 16):
            ws.cell(row=fila, column=col).value = None

    ws['A1'] = 'Checklist de Recepción de Mercancías'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:N1')
    ws['A2'] = ('Mes: _______________    '
                'Responsable: _________________________________')
    motor.cabecera(ws, 4, CAB_05, ANCHOS_05)

    for col in range(1, 15):
        cel = ws.cell(row=F0_05, column=col)
        cel.value = None
        if col in (2, 3, 14):
            motor.verde(cel, align='left')
        elif col == 4:
            motor.verde(cel, motor.FMT_TEMP)
        else:
            motor.verde(cel, '@')
    motor.replicar_filas(ws, F0_05, F0_05, F1_05, ncols=14, alto=20)

    motor.dv_lista(ws, '"✓,✗,N/A"',
                   [f'{c}{F0_05}:{c}{F1_05}' for c in 'FGHI'],
                   'Verificación',
                   'Marca ✓ (conforme), ✗ (no conforme) o N/A. Si escribes '
                   'otra cosa, el semáforo no la reconoce.')
    motor.dv_lista(ws, '"SÍ,NO"', [motor.rango('J', F0_05, F1_05)],
                   'Aceptado', 'Escribe SÍ o NO.')

    motor.semaforo(ws, f'F{F0_05}:I{F1_05}')
    motor.semaforo(ws, motor.rango('J', F0_05, F1_05),
                   extra_ok=('SÍ',), extra_rojo=('NO',))

    motor.sembrar(ws, 5, {'A': '05/09/2026', 'B': 'Merluza fresca',
                          'C': 'Pescados de la Ría S.L.', 'D': 1.5,
                          'E': '07/09/2026', 'F': '✓', 'G': '✓', 'H': '✓',
                          'I': '✓', 'J': 'SÍ', 'K': 'ALB-2026-4471',
                          'L': 'L-2026-0912', 'M': 'A.R.'}, marca_col='N')
    motor.sembrar(ws, 6, {'A': '05/09/2026', 'B': 'Canal de ternera',
                          'C': 'Cárnicas del Norte', 'D': 5.5,
                          'E': '12/09/2026', 'F': '✓', 'G': '✓', 'H': '✓',
                          'I': '✓', 'J': 'SÍ', 'K': 'ALB-2026-4472',
                          'L': 'L-2026-0913', 'M': 'A.R.'}, marca_col='N')
    motor.sembrar(ws, 7, {'A': '05/09/2026', 'B': 'Guisantes congelados',
                          'C': 'Congelados del Sur', 'D': -14,
                          'E': '03/2028', 'F': '✓', 'G': '✗', 'H': '✓',
                          'I': '✓', 'J': 'NO', 'K': 'ALB-2026-4473',
                          'L': 'L-2026-0914', 'M': 'A.R.',
                          'N': 'Envase roto y cadena de frío rota · '
                               'incidencia INC-002'}, marca_col='N')

    motor.nota(ws, F1_05 + 2,
               'Este checklist es la VERIFICACIÓN DOCUMENTAL de la entrega '
               '(albarán, lote, etiquetado, envase, cantidad y quién la '
               'recibió). El PCC de temperatura vive en el registro 02: no se '
               'fusionan porque acreditan cosas distintas y se archivan por '
               'separado. Anota el mismo nº de lote en los dos para que la '
               'trazabilidad (registro 06) los enganche.', ncols=14)
    motor.nota(ws, F1_05 + 3, 'Archivar junto con los albaranes de entrega. '
               + motor.CONSERVACION, ncols=14)
    motor.nota(ws, F1_05 + 4, motor.MARCA, ncols=14)
    motor.IMPRESION[(fname, ws.title)] = (4, True, 'C5')  # TEC-08

    motor.escribir_instrucciones(wb, 'Checklist de Recepción de Mercancías', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Verifica CADA entrega con este checklist y fírmalo: la columna '
              '«Firma receptor» es la que acredita quién hizo la comprobación.'),
        ('b', 'Anota el «Nº albarán» y el «Nº lote». Son las dos claves que '
              'enganchan esta hoja con el registro 06 de trazabilidad; sin '
              'ellas, una retirada de lote no se puede reconstruir.'),
        ('b', 'Las cuatro columnas de verificación admiten ✓, ✗ o N/A y se '
              'pintan solas (verde / rojo). Si escribes otra cosa, el '
              'desplegable la rechaza.'),
        ('b', 'Si algo no cumple: RECHAZAR la entrega y abrir la incidencia en '
              'el registro 11 (Acciones Correctivas).'),
        ('h', 'Qué se comprueba en cada entrega'),
        ('b', 'Temperatura: con termómetro de sonda verificado (registro 19). '
              'El veredicto por familia está en el registro 02.'),
        ('b', 'Caducidad: al menos 2/3 de vida útil restante.'),
        ('b', 'Etiquetado: lote, origen, ingredientes y alérgenos.'),
        ('b', 'Envase: sin golpes, roturas ni hinchazón.'),
        ('b', 'Aspecto y olor: sin signos de deterioro.'),
        ('h', 'Relación con el registro 02'),
        ('p', 'El 02 es el punto de control crítico de TEMPERATURA en '
              'recepción y el 05 es la verificación DOCUMENTAL de la misma '
              'entrega. No se fusionan: son dos registros distintos, con dos '
              'tarjetas distintas en tu panel, y una inspección puede pedir '
              'cualquiera de los dos por separado.'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Frecuencia: una fila por entrega. La hoja trae 40 filas '
              '(aprox. un mes).'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('05: albarán/lote/firma, 40 filas, DV que valida y '
                   'semáforo (DOM-23/TEC-14/TEC-26)')


# ==========================================================================
# 06 — Trazabilidad
# ==========================================================================
CAB_06 = ['Fecha entrada', 'Producto', 'Nº lote', 'Proveedor',
          'Fecha caducidad', 'Fecha consumo', 'Cantidad', 'Destino / uso',
          'Observaciones']
ANCHOS_06 = [15, 26, 16, 22, 15, 15, 12, 22, 28]
F0_06, F1_06 = 5, 44

CAB_06S = ['Fecha de salida', 'Lote de origen (→ hoja Trazabilidad)',
           'Elaboración / plato', 'Cantidad', 'Destino / servicio',
           'Responsable', 'Observaciones']
ANCHOS_06S = [15, 30, 28, 12, 26, 18, 28]
HOJA_SALIDA = 'Salida y uso interno'


def _post_06(wb, fname, cambios):
    ws = wb['Trazabilidad']
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for fila in range(F1_06 + 1, max(F1_06 + 8, ws.max_row + 1)):
        for col in range(1, 11):
            ws.cell(row=fila, column=col).value = None

    ws['A1'] = ('Registro de Trazabilidad — Entrada de productos '
                '(trazabilidad hacia atrás)')
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:I1')
    ws['A2'] = ('Mes: _______________    '
                'La trazabilidad hacia adelante va en la hoja '
                '«Salida y uso interno».')
    motor.cabecera(ws, 4, CAB_06, ANCHOS_06)

    for col in range(1, 10):
        cel = ws.cell(row=F0_06, column=col)
        cel.value = None
        motor.verde(cel, align='left' if col in (2, 4, 8, 9) else 'center')
    motor.replicar_filas(ws, F0_06, F0_06, F1_06, ncols=9, alto=20)

    motor.sembrar(ws, 5, {'A': '05/09/2026', 'B': 'Merluza fresca',
                          'C': 'L-2026-0912', 'D': 'Pescados de la Ría S.L.',
                          'E': '07/09/2026', 'F': '06/09/2026',
                          'G': '8,4 kg', 'H': 'Servicio de mediodía'},
                  marca_col='I')
    motor.sembrar(ws, 6, {'A': '05/09/2026', 'B': 'Canal de ternera',
                          'C': 'L-2026-0913', 'D': 'Cárnicas del Norte',
                          'E': '12/09/2026', 'F': '08/09/2026',
                          'G': '22 kg', 'H': 'Despiece y congelación'},
                  marca_col='I')
    motor.sembrar(ws, 7, {'A': '05/09/2026', 'B': 'Nata 35 % MG',
                          'C': 'L-2026-0915', 'D': 'Lácteos Vega',
                          'E': '30/09/2026', 'F': '10/09/2026',
                          'G': '6 L', 'H': 'Obrador — postres'},
                  marca_col='I')

    motor.nota(ws, F1_06 + 2,
               'En caso de alerta alimentaria: localizar el lote afectado en '
               'esta hoja, seguirlo en «Salida y uso interno» para saber a qué '
               'elaboraciones y servicios fue, retirar el producto y notificar '
               'de inmediato a la autoridad sanitaria de tu comunidad autónoma '
               '(art. 19 del Reg. (CE) 178/2002) y al proveedor.', ncols=9)
    motor.nota(ws, F1_06 + 3,
               'La información de trazabilidad debe poder entregarse de forma '
               'inmediata a requerimiento de la autoridad competente '
               '(Reg. (CE) 178/2002, art. 18). El reglamento no fija un plazo '
               'de «4 horas»: ese es un requisito de estándares privados tipo '
               'IFS/BRC.', ncols=9)
    motor.nota(ws, F1_06 + 4, motor.CONSERVACION, ncols=9)
    motor.nota(ws, F1_06 + 5, motor.MARCA, ncols=9)
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    # DOM-22/TEC-24: la trazabilidad hacia ADELANTE. Sin ella, en una alerta
    # real no se puede responder a la única pregunta que importa —a quién se
    # sirvió el lote— y el fichero contradecía a sus propias instrucciones y a
    # la tarjeta del panel, que prometen «entrada y salida».
    ss = motor.hoja(wb, HOJA_SALIDA)
    ss['A1'] = ('Registro de Trazabilidad — Salida y uso interno '
                '(trazabilidad hacia adelante)')
    ss['A1'].font = Font(bold=True, size=16)
    ss.merge_cells('A1:G1')
    ss['A2'] = ('Mes: _______________    Anota aquí en qué elaboración y a qué '
                'servicio fue cada lote de la hoja «Trazabilidad».')
    motor.cabecera(ss, 4, CAB_06S, ANCHOS_06S)
    for col in range(1, 8):
        cel = ss.cell(row=F0_06, column=col)
        cel.value = None
        motor.verde(cel, align='left' if col in (2, 3, 5, 7) else 'center')
    motor.replicar_filas(ss, F0_06, F0_06, F1_06, ncols=7, alto=20)
    motor.sembrar(ss, 5, {'A': '06/09/2026', 'B': 'L-2026-0912',
                          'C': 'Merluza a la bilbaína (12 raciones)',
                          'D': '4,2 kg', 'E': 'Servicio de mediodía — sala',
                          'F': 'A.R.'}, marca_col='G')
    motor.sembrar(ss, 6, {'A': '08/09/2026', 'B': 'L-2026-0913',
                          'C': 'Estofado de ternera (tanda de 30 raciones)',
                          'D': '9 kg', 'E': 'Catering empresa Nordel',
                          'F': 'M.S.'}, marca_col='G')
    motor.nota(ss, F1_06 + 2,
               'Una salida por línea. Si un lote se reparte entre varias '
               'elaboraciones o servicios, usa una línea por cada uno: es lo '
               'que permite acotar la retirada a lo que de verdad salió.',
               ncols=7)
    motor.nota(ss, F1_06 + 3, motor.CONSERVACION, ncols=7)
    motor.nota(ss, F1_06 + 4, motor.MARCA, ncols=7)
    motor.IMPRESION[(fname, HOJA_SALIDA)] = (4, True)

    motor.escribir_instrucciones(wb, 'Registro de Trazabilidad', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Hoja «Trazabilidad»: TODO lo que ENTRA, con su lote y su '
              'proveedor. Es la trazabilidad hacia atrás.'),
        ('b', 'Hoja «Salida y uso interno»: en qué elaboración y a qué '
              'servicio fue cada lote. Es la trazabilidad hacia adelante, y es '
              'la que responde a «¿a quién se sirvió esto?» en una alerta.'),
        ('b', 'Anota siempre lote, proveedor, fecha de entrada y fecha de '
              'caducidad. El mismo nº de lote debe aparecer en el registro 05 '
              '(recepción) para que las dos hojas se enganchen.'),
        ('b', 'Obligatorio según el Reglamento (CE) 178/2002, arts. 18 y 19.'),
        ('h', 'Qué exige realmente el Reg. (CE) 178/2002'),
        ('b', 'Art. 18: poder identificar a quién se compró y a quién se '
              'suministró cada producto, y entregar esa información de forma '
              'inmediata a requerimiento de la autoridad competente. NO fija '
              'un plazo de «4 horas»: eso es de estándares privados IFS/BRC.'),
        ('b', 'Art. 19: si sospechas que un alimento que has puesto en el '
              'mercado puede ser nocivo, retirarlo e informar de inmediato a '
              'la autoridad sanitaria de tu comunidad autónoma.'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Frecuencia: una línea por entrada y una línea por salida. '
              '40 filas por hoja (aprox. un mes).'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('06: pestaña «Salida y uso interno», «< 4 horas» corregido '
                   'y conservación unificada (DOM-22/TEC-24/COM-21/DOM-35)')


# ==========================================================================
# 09 — Control de aceite de fritura
# ==========================================================================
CAB_09 = ['Fecha', 'Freidora / equipo', 'Tipo de test', 'Resultado (% CP)',
          'Temp. máx. alcanzada (°C)', 'Estado', 'Acción realizada', 'Firma']
ANCHOS_09 = [14, 20, 20, 16, 20, 14, 34, 16]
F0_09, F1_09 = 5, 44
HOJA_RETIRADA = 'Retirada de aceite usado'
CAB_09R = ['Fecha', 'Litros retirados', 'Gestor autorizado',
           'Nº autorización / documento', 'Firma', 'Observaciones']
ANCHOS_09R = [14, 16, 34, 30, 16, 26]


def _post_09(wb, fname, cambios):
    ws = wb['Control Aceite']
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for fila in range(F1_09 + 1, max(F1_09 + 8, ws.max_row + 1)):
        for col in range(1, 10):
            ws.cell(row=fila, column=col).value = None

    ws['A1'] = 'Control de Aceite de Fritura'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A2'] = ('Freidora: ________________________    '
                'Tipo de aceite: ________________________')
    motor.cabecera(ws, 4, CAB_09, ANCHOS_09)

    for col in range(1, 9):
        cel = ws.cell(row=F0_09, column=col)
        cel.value = None
        if col == 6:
            motor.calculada(cel)
        elif col in (4, 5):
            motor.verde(cel, motor.FMT_TEMP)
        elif col in (2, 3, 7):
            motor.verde(cel, align='left')
        else:
            motor.verde(cel, '@')

    # TEC-06: el umbral era `>25`, así que 25,0 % clavado —la lectura que da un
    # medidor digital, que devuelve enteros— salía VIGILAR mientras el pie del
    # propio fichero ordenaba cambiar ANTES de superar el límite.
    # TEC-07/DOM-25/COM-19: la temperatura era una columna muerta. Se puede
    # freír a 210 °C y, con el %CP bajo, la fila decía OK. Ahora entra en el
    # veredicto. `IF(E="",0,E)` evita comparar un vacío con 180.
    # DOM-R2-08 / TEC-10 (ronda 2): la guarda miraba SÓLO el %CP, así que una
    # freidora anotada a 220 o 240 °C sin test se quedaba con el Estado en
    # BLANCO —sin semáforo, indistinguible de una fila conforme en la hoja
    # impresa— pese a que el pie del propio fichero promete lo contrario. Se
    # invierte igual que en el 10 (DOM-32/TEC-15): la temperatura manda aunque
    # no hayas llegado a hacer el test, y la falta del test se declara.
    f = (f'=IF(AND($D{F0_09}="",$E{F0_09}=""),"",'
         f'IF(IF($E{F0_09}="",0,$E{F0_09})>180,"CAMBIAR",'
         f'IF($D{F0_09}="","FALTA TEST",'
         f'IF($D{F0_09}>=25,"CAMBIAR",'
         f'IF($D{F0_09}>=20,"VIGILAR","OK")))))')
    ws.cell(row=F0_09, column=6).value = f
    motor.reg(ws, f'F{F0_09}', f)
    motor.replicar_filas(ws, F0_09, F0_09, F1_09, ncols=8, alto=20)

    motor.dv_lista(ws, '"Tiras reactivas,Medidor digital,Visual,Otro"',
                   [motor.rango('C', F0_09, F1_09)], 'Tipo de test',
                   'Elige un método de la lista.')
    motor.dv_decimal(ws, [motor.rango('D', F0_09, F1_09)],
                     'Compuestos polares',
                     'Introduce el % de compuestos polares, entre 0 y 50.',
                     minimo=0, maximo=50)
    motor.dv_decimal(ws, [motor.rango('E', F0_09, F1_09)],
                     'Temperatura máxima',
                     'Introduce la temperatura máxima alcanzada en °C, entre '
                     '0 y 250.', minimo=0, maximo=250)
    motor.semaforo(ws, motor.rango('F', F0_09, F1_09),
                   extra_ambar=('FALTA TEST',))

    motor.sembrar(ws, 5, {'A': '01/09/2026', 'B': 'Freidora 1',
                          'C': 'Tiras reactivas', 'D': 12, 'E': 175,
                          'G': 'Filtrado y reposición de nivel', 'H': 'A.R.'},
                  marca_col='G')
    motor.sembrar(ws, 6, {'A': '08/09/2026', 'B': 'Freidora 1',
                          'C': 'Medidor digital', 'D': 21, 'E': 178,
                          'G': 'Vigilar: previsto cambio esta semana',
                          'H': 'A.R.'}, marca_col='G')
    motor.sembrar(ws, 7, {'A': '15/09/2026', 'B': 'Freidora 2',
                          'C': 'Medidor digital', 'D': 26, 'E': 180,
                          'G': 'Cambio de aceite y limpieza de cuba',
                          'H': 'M.S.'}, marca_col='G')

    motor.nota(ws, F1_09 + 2,
               'Límite legal: máx. 25 % de compuestos polares — Orden de 26 de '
               'enero de 1989, por la que se aprueba la Norma de Calidad para '
               'los aceites y grasas calentados. Cambiar el aceite al alcanzar '
               'el límite, no después: por eso 25,0 % ya dice CAMBIAR.',
               ncols=8)
    motor.nota(ws, F1_09 + 3,
               'Temperatura máxima de fritura recomendada: 180 °C. Por encima '
               'la degradación se dispara, así que el Estado también dice '
               'CAMBIAR aunque el % de polares sea bajo o no se haya medido.',
               ncols=8)
    motor.nota(ws, F1_09 + 4,
               'Si anotas la temperatura y todavía no has hecho el test de '
               'compuestos polares, el Estado dice FALTA TEST en ámbar: la '
               'fila no acredita nada hasta que lo completes.', ncols=8)
    motor.nota(ws, F1_09 + 5, motor.CONSERVACION, ncols=8)
    motor.nota(ws, F1_09 + 6, motor.MARCA, ncols=8)
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    rs = motor.hoja(wb, HOJA_RETIRADA)
    rs['A1'] = 'Retirada de Aceite Usado por Gestor Autorizado'
    rs['A1'].font = Font(bold=True, size=16)
    rs.merge_cells('A1:F1')
    rs['A2'] = ('El aceite vegetal usado es un residuo no peligroso (LER '
                '20 01 25): sólo puede retirarlo un gestor autorizado, y el '
                'justificante se pide en inspección junto al control de '
                'fritura.')
    motor.cabecera(rs, 4, CAB_09R, ANCHOS_09R)
    for col in range(1, 7):
        cel = rs.cell(row=5, column=col)
        cel.value = None
        motor.verde(cel, align='left' if col in (3, 4, 6) else 'center')
    motor.replicar_filas(rs, 5, 5, 28, ncols=6, alto=20)
    motor.sembrar(rs, 5, {'A': '15/09/2026', 'B': 40,
                          'C': 'Recogidas Oleo S.L.',
                          'D': 'RET-2026-0915', 'E': 'M.S.'}, marca_col='F')
    motor.nota(rs, 30, 'Guarda el justificante de cada retirada junto con este '
                       'registro. ' + motor.CONSERVACION, ncols=6)
    motor.nota(rs, 31, motor.MARCA, ncols=6)
    motor.IMPRESION[(fname, HOJA_RETIRADA)] = (4, True)

    motor.escribir_instrucciones(wb, 'Control de Aceite de Fritura', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Registra cada test de aceite en la hoja «Control Aceite»: '
              'fecha, freidora, método, % de compuestos polares y temperatura '
              'máxima alcanzada ese día.'),
        ('b', 'La columna «Estado» se calcula sola y se pinta: OK (verde), '
              'VIGILAR (ámbar) a partir del 20 % y CAMBIAR (rojo) al alcanzar '
              'el 25 % O al superar los 180 °C.'),
        ('b', 'Mide con tiras reactivas o con medidor digital de compuestos '
              'polares. El método «Visual» no es un control: úsalo sólo como '
              'apoyo.'),
        ('b', 'Anota cada retirada de aceite usado en la hoja «Retirada de '
              'aceite usado», con el gestor y el nº de documento.'),
        ('h', 'Límites que aplica la hoja'),
        ('b', 'Máximo 25 % de compuestos polares — Orden de 26 de enero de '
              '1989 (Norma de Calidad para los aceites y grasas calentados). '
              'El 25,0 % clavado ya obliga a cambiar.'),
        ('b', 'Temperatura máxima de fritura recomendada: 180 °C. Por encima, '
              'el Estado dice CAMBIAR aunque el % de polares sea bajo o '
              'todavía no lo hayas medido.'),
        ('b', 'Si anotas sólo la temperatura y aún no has hecho el test, el '
              'Estado dice FALTA TEST en ámbar. La fila deja de estar en '
              'blanco: se ve que falta el dato, no que todo esté bien.'),
        ('b', 'El aceite usado es un residuo (LER 20 01 25): sólo lo retira un '
              'gestor autorizado y hay que guardar el justificante.'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Frecuencia recomendada: un test por freidora y semana en uso '
              'normal; diario si se fríe empanado o pescado a diario. La hoja '
              'trae 40 filas.'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('09: umbral en 25,0 %, temperatura dentro del veredicto, '
                   'guarda invertida con FALTA TEST y pestaña de retirada '
                   '(TEC-06/TEC-07/DOM-25/COM-19/DOM-R2-08/TEC-10)')


# ==========================================================================
# 10 — Control de agua potable
# ==========================================================================
CAB_10 = ['Fecha', 'Punto de muestreo', 'Cloro residual (mg/L)',
          'Aspecto / olor', 'Estado', 'Análisis externo (S/N)',
          'Observaciones']
ANCHOS_10 = [14, 22, 20, 20, 16, 18, 32]
F0_10, F1_10 = 5, 35                              # 31 filas (COM-24)


def _post_10(wb, fname, cambios):
    ws = wb['Control Agua']
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for fila in range(F1_10 + 1, max(F1_10 + 8, ws.max_row + 1)):
        for col in range(1, 9):
            ws.cell(row=fila, column=col).value = None

    ws['A1'] = 'Registro de Control de Agua Potable'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')
    ws['A2'] = ('Fuente de suministro: ☐ Red municipal   ☐ Pozo propio   '
                '☐ Depósito        Responsable: ____________________')
    motor.cabecera(ws, 4, CAB_10, ANCHOS_10)

    for col in range(1, 8):
        cel = ws.cell(row=F0_10, column=col)
        cel.value = None
        if col == 5:
            motor.calculada(cel)
        elif col == 3:
            motor.verde(cel, '0.00')
        elif col in (2, 7):
            motor.verde(cel, align='left')
        else:
            motor.verde(cel, '@')

    # DOM-32/TEC-15: las dos guardas estaban invertidas. Con el cloro correcto
    # y el aspecto sin rellenar salía REVISAR (falsa no conformidad); con el
    # aspecto TURBIO y el cloro sin medir —el caso típico: se ve el problema
    # antes de sacar el kit— salía cadena vacía, y el único indicio registrado
    # de agua no potable no disparaba nada. Ahora el aspecto anormal manda
    # SIEMPRE, y la falta de un dato se declara como tal.
    f = (f'=IF(AND($C{F0_10}="",$D{F0_10}=""),"",'
         f'IF(AND($D{F0_10}<>"",$D{F0_10}<>"Normal"),"REVISAR",'
         f'IF($C{F0_10}="","FALTA CLORO",'
         f'IF($D{F0_10}="","INCOMPLETO",'
         f'IF(AND($C{F0_10}>=0.2,$C{F0_10}<=1),"OK","REVISAR")))))')
    ws.cell(row=F0_10, column=5).value = f
    motor.reg(ws, f'E{F0_10}', f)
    motor.replicar_filas(ws, F0_10, F0_10, F1_10, ncols=7, alto=20)

    motor.dv_lista(ws, '"Normal,Turbio,Olor extraño,Color anormal"',
                   [motor.rango('D', F0_10, F1_10)], 'Aspecto / olor',
                   'Elige una opción de la lista. El Estado la compara con '
                   '«Normal», así que un texto libre no se reconoce.')
    motor.dv_lista(ws, '"S,N"', [motor.rango('F', F0_10, F1_10)],
                   'Análisis externo', 'Escribe S o N.')
    motor.dv_decimal(ws, [motor.rango('C', F0_10, F1_10)], 'Cloro residual',
                     'Introduce el cloro residual libre en mg/L, entre 0 y 5.',
                     minimo=0, maximo=5)
    # FALTA CLORO va en ÁMBAR: no es una no conformidad del agua, es un dato
    # que falta. Distinguirlo de REVISAR es justo lo que pedía TEC-15.
    motor.semaforo(ws, motor.rango('E', F0_10, F1_10),
                   extra_ambar=('FALTA CLORO',))

    motor.sembrar(ws, 5, {'A': '01/09/2026', 'B': 'Grifo cocina', 'C': 0.6,
                          'D': 'Normal', 'F': 'N'}, marca_col='G')
    motor.sembrar(ws, 6, {'A': '08/09/2026', 'B': 'Grifo office', 'C': 0.15,
                          'D': 'Normal', 'F': 'N',
                          'G': 'Cloro por debajo de 0,2 mg/L: avisado el '
                               'gestor de la red'}, marca_col='G')
    motor.sembrar(ws, 7, {'A': '15/09/2026', 'B': 'Grifo cocina',
                          'D': 'Turbio', 'F': 'S',
                          'G': 'Agua turbia: cerrado el punto y solicitado '
                               'análisis externo · incidencia INC-003'},
                  marca_col='G')

    motor.nota(ws, F1_10 + 2,
               'Cloro residual libre aceptable en el punto de consumo: '
               '0,2 - 1,0 mg/L (RD 3/2023, de 10 de enero, sobre criterios '
               'técnico-sanitarios de la calidad del agua de consumo).',
               ncols=7)
    motor.nota(ws, F1_10 + 3,
               'INCOMPLETO = falta el aspecto. FALTA CLORO = falta la medida '
               'de cloro. REVISAR = hay una desviación real. La hoja nunca '
               'da OK con un dato sin rellenar.', ncols=7)
    motor.nota(ws, F1_10 + 4, motor.CONSERVACION, ncols=7)
    motor.nota(ws, F1_10 + 5, motor.MARCA, ncols=7)
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    motor.escribir_instrucciones(wb, 'Registro de Control de Agua Potable', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Marca arriba de qué fuente viene tu agua: red municipal, pozo '
              'propio o depósito. De eso depende la frecuencia.'),
        ('b', 'Anota el cloro residual libre en mg/L y elige el aspecto / olor '
              'en el desplegable. El Estado se calcula solo y se pinta.'),
        ('b', 'Si tienes pozo o depósito propio, además son obligatorios los '
              'análisis por laboratorio: marca «Análisis externo = S» y guarda '
              'el boletín con esta hoja.'),
        ('b', 'Si el agua viene de la red municipal, guarda también los '
              'boletines que publica tu gestor de abastecimiento.'),
        ('h', 'Qué significa cada Estado'),
        ('b', 'OK — cloro entre 0,2 y 1,0 mg/L y aspecto Normal.'),
        ('b', 'INCOMPLETO (ámbar) — has medido el cloro pero falta el aspecto.'),
        ('b', 'FALTA CLORO (ámbar) — el aspecto es normal pero no hay medida '
              'de cloro.'),
        ('b', 'REVISAR (rojo) — el aspecto es anormal (turbio, olor extraño, '
              'color) o el cloro está fuera de rango. El aspecto anormal manda '
              'aunque no hayas llegado a medir el cloro.'),
        ('h', 'Frecuencia recomendada'),
        ('b', 'Con depósito propio o pozo: control diario de cloro residual en '
              'un punto de la instalación (por eso la hoja trae 31 filas: un '
              'mes completo).'),
        ('b', 'Con agua de red y sin depósito: control semanal, rotando los '
              'puntos de muestreo.'),
        ('b', 'Tras cualquier obra, limpieza de depósito o corte de suministro: '
              'medir antes de volver a usar el agua.'),
        ('h', 'Marco normativo y archivo'),
        ('b', 'RD 3/2023, de 10 de enero, por el que se establecen los '
              'criterios técnico-sanitarios de la calidad del agua de consumo '
              '(sustituye íntegramente al reglamento anterior).'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('10: guardas invertidas con INCOMPLETO/FALTA CLORO, '
                   '31 filas y frecuencia documentada (DOM-32/TEC-15/COM-24)')


# ==========================================================================
# API del grupo
# ==========================================================================
PRE = {
    '02-registro-temperaturas-recepcion.xlsx': _pre_02,
    '05-checklist-recepcion-mercancias.xlsx': _pre_05,
}
POST = {
    '01-registro-temperaturas-diario.xlsx': _post_01,
    '02-registro-temperaturas-recepcion.xlsx': _post_02,
    '05-checklist-recepcion-mercancias.xlsx': _post_05,
    '06-registro-trazabilidad.xlsx': _post_06,
    '09-control-aceite-fritura.xlsx': _post_09,
    '10-control-agua-potable.xlsx': _post_10,
}


def pre(wb, fname, cambios):
    fn = PRE.get(fname)
    if fn:
        fn(wb, cambios)


def post(wb, fname, cambios, registro=None):
    fn = POST.get(fname)
    if fn:
        fn(wb, fname, cambios)


# ==========================================================================
# §6 — lo que hay que poder demostrar en la ronda de refutación
# ==========================================================================
def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}: {e}'


def _caso(carpeta, destino, fname, hoja, entradas, salida, esperado, lectura):
    """Escribe `entradas` en una COPIA desechable y evalúa `salida` con pycel.

    Nunca sobre el entregable: la demostración no puede dejar datos de prueba
    en el fichero que se publica.
    """
    from pycel import ExcelCompiler
    os.makedirs(destino, exist_ok=True)
    etiqueta = f"{fname.split('-')[0]}-{salida.replace('!', '_').replace(chr(39), '')}"
    dst = os.path.join(destino, f'demo-{etiqueta}.xlsx')
    shutil.copy2(os.path.join(carpeta, fname), dst)
    import openpyxl
    wb = openpyxl.load_workbook(dst)
    ws = wb[hoja]
    for coord, valor in entradas.items():
        ws[coord] = valor
    wb.save(dst)
    xl = ExcelCompiler(filename=dst)
    obtenido = _ev(xl, f"'{hoja}'!{salida}")
    return {
        'ref': f'{fname}:{hoja}:{salida}',
        'entradas': {f'{hoja}!{k}': v for k, v in entradas.items()},
        'esperado': esperado, 'obtenido': obtenido,
        'ok': obtenido == esperado, 'lectura': lectura,
        'copia_desechable': dst,
    }


def demos(carpeta, origen, destino):
    """Los siete casos del §6 que tocan al grupo A, más el barrido de normas
    derogadas."""
    import openpyxl
    casos = []
    fam_pescado = lim_02('Pescado fresco')
    fam_vacuno = lim_02('Canales')

    casos.append(_caso(
        carpeta, destino, '01-registro-temperaturas-diario.xlsx',
        'Registro Semanal', {'B27': -26}, 'C27', 'OK',
        'Arcón a -26 °C: conforme. En v1.1 el rango cerrado -25/-18 lo marcaba '
        'ALERTA y el pie de hoja mandaba abrir una acción correctiva '
        '(DOM-09/TEC-05/COM-18).'))
    casos.append(_caso(
        carpeta, destino, '01-registro-temperaturas-diario.xlsx',
        'Registro Semanal', {'B57': 102}, 'C57', 'OK',
        'Baño maría a 102 °C: conforme. En v1.1 el techo de 100 °C lo marcaba '
        'ALERTA (TEC-31).'))
    casos.append(_caso(
        carpeta, destino, '02-registro-temperaturas-recepcion.xlsx',
        'Recepción Temperaturas', {'D20': 4, 'E20': fam_pescado}, 'G20',
        'RECHAZAR',
        'Pescado fresco a 4 °C: se rechaza. Debe recibirse en hielo fundente '
        '(0-2 °C); el desplegable único de 4 °C de v1.1 lo daba por conforme '
        '(DOM-04).'))
    casos.append(_caso(
        carpeta, destino, '02-registro-temperaturas-recepcion.xlsx',
        'Recepción Temperaturas', {'D21': 6, 'E21': fam_vacuno}, 'G21', 'OK',
        'Canal de vacuno a 6 °C: conforme (límite legal 7 °C). En v1.1 se '
        'rechazaba al proveedor sin motivo (DOM-04).'))
    casos.append(_caso(
        carpeta, destino, '02-registro-temperaturas-recepcion.xlsx',
        'Recepción Temperaturas', {'D22': -5}, 'G22', '',
        'Temperatura anotada SIN elegir familia: el Estado se queda en blanco. '
        'En v1.1 el límite vacío se coaccionaba a 0 y un congelado a -5 °C '
        'salía OK (TEC-01, alta).'))
    casos.append(_caso(
        carpeta, destino, '09-control-aceite-fritura.xlsx', 'Control Aceite',
        {'D20': 25.0}, 'F20', 'CAMBIAR',
        '25,0 % de compuestos polares clavado: CAMBIAR. En v1.1 el umbral era '
        '«> 25» y decía VIGILAR, contradiciendo al pie del propio fichero '
        '(TEC-06).'))
    casos.append(_caso(
        carpeta, destino, '09-control-aceite-fritura.xlsx', 'Control Aceite',
        {'D21': 18, 'E21': 200}, 'F21', 'CAMBIAR',
        '18 % de polares pero 200 °C: CAMBIAR. En v1.1 la columna de '
        'temperatura no la leía ninguna fórmula (TEC-07/DOM-25/COM-19).'))
    casos.append(_caso(
        carpeta, destino, '10-control-agua-potable.xlsx', 'Control Agua',
        {'D20': 'Turbio'}, 'E20', 'REVISAR',
        'Agua turbia sin medida de cloro: REVISAR. En v1.1 la guarda sólo '
        'miraba el cloro y devolvía cadena vacía (TEC-15).'))
    casos.append(_caso(
        carpeta, destino, '10-control-agua-potable.xlsx', 'Control Agua',
        {'C21': 0.6}, 'E21', 'INCOMPLETO',
        'Cloro correcto y aspecto sin rellenar: INCOMPLETO, no REVISAR. En '
        'v1.1 era una falsa no conformidad indistinguible de una real '
        '(DOM-32).'))

    # §6: «RD 2207/1995», «RD 140/2003», «carné de manipulador» y «€60.000»
    # con 0 ocurrencias en los ficheros del grupo.
    normativa = []
    for fname in FICHEROS:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for etiqueta, rx in motor.PROHIBIDAS:
            donde = [f'{ws.title}!{c.coordinate}'
                     for ws in wb.worksheets for row in ws.iter_rows()
                     for c in row
                     if isinstance(c.value, str) and rx.search(c.value)]
            normativa.append({'fichero': fname, 'cadena': etiqueta,
                              'patron': rx.pattern,
                              'ocurrencias': len(donde), 'donde': donde})

    fallos = [f"§6 {c['ref']}: esperaba {c['esperado']!r}, dio {c['obtenido']!r}"
              for c in casos if not c['ok']]
    fallos += [f"§6 «{n['cadena']}» sigue viva en {n['fichero']}: {n['donde']}"
               for n in normativa if n['ocurrencias']]

    return {'fallos': fallos,
            'casos_spec_6': casos,
            'normativa_derogada': normativa,
            'normativa_ocurrencias_totales':
                sum(n['ocurrencias'] for n in normativa)}


# ==========================================================================
# Casos «dato FUERA de límite» que consume `main.py` (paso 5).
# Complementan a §6: allí se demuestra que lo CONFORME deja de dar falsas
# alarmas; aquí, que lo NO conforme sigue disparando. Las dos mitades hacen
# falta: una fórmula que devolviera «OK» siempre pasaría la primera.
# ==========================================================================
CASOS_LIMITE = [
    {'fichero': '01-registro-temperaturas-diario.xlsx',
     'hoja': 'Registro Semanal', 'entradas': {'B10': 9.5}, 'salida': 'C10',
     'esperado': 'ALERTA', 'lectura': 'Cámara de refrigeración a 9,5 °C'},
    {'fichero': '01-registro-temperaturas-diario.xlsx',
     'hoja': 'Registro Semanal', 'entradas': {'G30': -10}, 'salida': 'H30',
     'esperado': 'ALERTA', 'lectura': 'Congelador a -10 °C en la lectura de la '
                                      'tarde (por encima de -18)'},
    {'fichero': '01-registro-temperaturas-diario.xlsx',
     'hoja': 'Registro Semanal', 'entradas': {'B60': 55}, 'salida': 'C60',
     'esperado': 'ALERTA', 'lectura': 'Baño maría a 55 °C (por debajo de 65)'},
    {'fichero': '02-registro-temperaturas-recepcion.xlsx',
     'hoja': 'Recepción Temperaturas',
     'entradas': {'D25': 8, 'E25': lim_02('Canales')}, 'salida': 'G25',
     'esperado': 'RECHAZAR', 'lectura': 'Canal de vacuno a 8 °C (límite 7)'},
    {'fichero': '02-registro-temperaturas-recepcion.xlsx',
     'hoja': 'Recepción Temperaturas',
     'entradas': {'D26': -5, 'E26': lim_02('Congelados')}, 'salida': 'G26',
     'esperado': 'RECHAZAR', 'lectura': 'Congelado recibido a -5 °C'},
    {'fichero': '09-control-aceite-fritura.xlsx', 'hoja': 'Control Aceite',
     'entradas': {'D25': 30}, 'salida': 'F25', 'esperado': 'CAMBIAR',
     'lectura': '30 % de compuestos polares'},
    {'fichero': '09-control-aceite-fritura.xlsx', 'hoja': 'Control Aceite',
     'entradas': {'D25': 22}, 'salida': 'F25', 'esperado': 'VIGILAR',
     'lectura': '22 % de compuestos polares (tramo ámbar)'},
    # DOM-R2-08 / TEC-10: la guarda invertida. Antes las dos filas se quedaban
    # con el Estado en blanco.
    {'fichero': '09-control-aceite-fritura.xlsx', 'hoja': 'Control Aceite',
     'entradas': {'E26': 220}, 'salida': 'F26', 'esperado': 'CAMBIAR',
     'lectura': 'Freidora anotada a 220 °C sin test de polares: CAMBIAR '
                '(antes se quedaba en blanco)'},
    {'fichero': '09-control-aceite-fritura.xlsx', 'hoja': 'Control Aceite',
     'entradas': {'E27': 170}, 'salida': 'F27', 'esperado': 'FALTA TEST',
     'lectura': 'Temperatura correcta pero sin test de polares: FALTA TEST en '
                'ámbar, no OK ni blanco'},
    # DOM-R2-06 / DOM-R2-22: familias partidas según el Reg. (CE) 853/2004.
    {'fichero': '02-registro-temperaturas-recepcion.xlsx',
     'hoja': 'Recepción Temperaturas',
     'entradas': {'D27': 3, 'E27': lim_02('Carne picada')}, 'salida': 'G27',
     'esperado': 'RECHAZAR', 'lectura': 'Carne picada a 3 °C: RECHAZAR '
                                        '(límite 2 °C, Secc. V Cap. III)'},
    {'fichero': '02-registro-temperaturas-recepcion.xlsx',
     'hoja': 'Recepción Temperaturas',
     'entradas': {'D28': 3, 'E28': lim_02('Preparados de carne')},
     'salida': 'G28', 'esperado': 'OK',
     'lectura': 'Preparado de carne (hamburguesa) a 3 °C: OK. Con la familia '
                'sin partir salía RECHAZAR sobre una entrega legal '
                '(DOM-R2-06)'},
    {'fichero': '02-registro-temperaturas-recepcion.xlsx',
     'hoja': 'Recepción Temperaturas',
     'entradas': {'D29': 6, 'E29': lim_02('Caza mayor')}, 'salida': 'G29',
     'esperado': 'OK',
     'lectura': 'Corzo (caza mayor silvestre) a 6 °C: OK. Antes iba con las '
                'aves a 4 °C y se rechazaba (DOM-R2-22)'},
    {'fichero': '10-control-agua-potable.xlsx', 'hoja': 'Control Agua',
     'entradas': {'C25': 0.05, 'D25': 'Normal'}, 'salida': 'E25',
     'esperado': 'REVISAR', 'lectura': 'Cloro residual 0,05 mg/L'},
]
