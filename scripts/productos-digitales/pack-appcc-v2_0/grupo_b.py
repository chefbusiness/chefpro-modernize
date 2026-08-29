#!/usr/bin/env python3
"""
grupo_b.py — §3 de `pack-appcc-v2-SPEC.md`: planes, checklists, HACCP y guía
(03, 04, 07, 11, 12, 13, 14, 15).

Al contrario que el grupo A —seis registros con fórmula de veredicto—, aquí lo
que falla es el CONTENIDO que se enseña a un inspector: un plan de limpieza que
prescribe mezclar ácido y lejía, un análisis de peligros que remite a registros
que el pack no entrega, una escala de sanciones inventada y una guía que promete
25 puntos y trae 23. Por eso este grupo escribe sobre todo texto defendible, y
sólo pone fórmula donde hay algo que contar.

  · 03 plan L+D — fregaderos sin ácido+lejía (DOM-05, alta), bloque EXTERIOR y
    los elementos que faltaban hasta 32 filas (DOM-11/TEC-12/COM-06), columnas
    «Tiempo de contacto» y «Nº registro / FDS» + pestaña de productos químicos
    (DOM-12), DV de frecuencia ampliada y C9 partida (DOM-31/TEC-11).
  · 04 limpieza diaria — 14 columnas M/T por día y bloque de TAREAS SEMANALES
    (DOM-13/TEC-27), DV «✓,✗,N/A» sobre las casillas (TEC-22) y print_area
    hasta el pie (DOM-29/TEC-13/COM-23).
  · 07 plagas — Nº ROESB en cabecera, «Nº registro del biocida» y «Plazo de
    seguridad», pestaña «Plano de cebos» (DOM-16) y 80 filas (TEC-20).
  · 11 acciones correctivas — producto/lote afectado, destino del producto no
    conforme, acción preventiva y registro de origen (DOM-24).
  · 12 análisis de peligros — medida preventiva, vigilancia y verificación
    (DOM-17), nivel de riesgo CALCULADO (TEC-08), filas 6 y 17 reclasificadas
    como PPRo (COM-20), descongelación, huevo fresco y anisakis (DOM-06/
    DOM-26), la columna «Registro» reapuntada a ficheros que existen
    (DOM-01/TEC-04/COM-04) y DV sobre TODAS las filas (TEC-25).
  · 13 higiene personal — formación acreditada en vez del carné suprimido
    (DOM-14) y área de impresión que no corta la línea de firma (TEC-29).
  · 14 fichas de alérgenos — protocolo de 10 pasos en dos bloques (DOM-27) con
    wrap y altura (TEC-09), y los sulfitos «expresado como SO2» (TEC-32).
  · 15 guía de inspección — 25 puntos REALES (DOM-10/COM-07), escala de la Ley
    17/2011 (DOM-28), bloques de preparación, documentos y errores, y resumen
    con COUNTIFS de muy graves y graves, % de cumplimiento y «Sin responder»
    (TEC-16/TEC-10).

Igual que `grupo_a.py`, se ejecuta alrededor del motor:
  pre(wb, fname, cambios)   → inserciones estructurales. Aquí NO hace falta
                              ninguna: cada hoja se reconstruye entera desde
                              cero, que es la forma más barata de ser
                              idempotente (la 2.ª pasada escribe exactamente lo
                              mismo en las mismas celdas).
  post(wb, fname, cambios, registro) → rejilla, fórmulas, DV, semáforo,
                              ejemplos, pies e Instrucciones.
"""
import contextlib
import os
import shutil

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter

import motor

FICHEROS = [
    '03-plan-limpieza-desinfeccion.xlsx',
    '04-registro-limpieza-diaria.xlsx',
    '07-control-plagas-ddd.xlsx',
    '11-registro-acciones-correctivas.xlsx',
    '12-analisis-peligros-haccp.xlsx',
    '13-checklist-higiene-personal.xlsx',
    '14-fichas-14-alergenos.xlsx',
    '15-guia-inspeccion-sanidad.xlsx',
]


# ==========================================================================
# Utilidades locales
# ==========================================================================
def _lienzo(ws, filas, cols):
    """Deja la hoja en blanco: sin merges, sin valores, sin estilos y sin
    alturas ni anchos heredados.

    `motor.hoja()` limpia valores, merges, DV y CF pero NO estilos: en la 2.ª
    pasada eso no rompe la idempotencia (el digest sólo mira celdas con valor)
    pero sí deja restos visibles — el relleno verde de las doce filas vacías
    que el 03 arrastraba de v1.1, por ejemplo. Aquí se borra todo y se pinta
    desde cero.
    """
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    nf = max(filas, ws.max_row)
    nc = max(cols, ws.max_column)
    for r in range(1, nf + 1):
        for c in range(1, nc + 1):
            cel = ws.cell(row=r, column=c)
            cel.value = None
            cel._style = StyleArray()
    # Se BORRAN las entradas, no se ponen a None: `ColumnDimension.width` es un
    # Float que no admite None, y dejar la entrada vacía escribiría un <col>
    # inútil en el XML que además rompería la comparación de idempotencia.
    for k in list(ws.row_dimensions):
        del ws.row_dimensions[k]
    for k in list(ws.column_dimensions):
        del ws.column_dimensions[k]


def _titulo(ws, texto, ncols, sub=None, sub2=None):
    ws['A1'] = texto
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    if sub:
        ws['A2'] = sub
        ws['A2'].font = Font(size=10)
    if sub2:
        ws['A3'] = sub2
        ws['A3'].font = Font(size=10, italic=True)


def _rangos_filas(col, filas):
    """['C6:C21', 'C23:C26', …] a partir de una lista de filas sueltas.

    Hace falta porque las bandas de sección parten la tabla: una DV o un
    semáforo con varias áreas en el mismo sqref se evalúa relativo al primer
    vértice y las demás salen desplazadas (ver `motor.semaforo`).
    """
    out, ini, prev = [], None, None
    for f in sorted(filas):
        if ini is None:
            ini = prev = f
        elif f == prev + 1:
            prev = f
        else:
            out.append(f'{col}{ini}:{col}{prev}')
            ini = prev = f
    if ini is not None:
        out.append(f'{col}{ini}:{col}{prev}')
    return out


def _alto(texto, ancho_car, base=14, minimo=18):
    n = max(1, -(-len(str(texto)) // max(10, int(ancho_car * 1.05))))
    return max(minimo, base * n + 4)


def _contador(ws, fila, ncols, col_valor, etiqueta, formula, fmt=None,
              rojo_si=None):
    """Etiqueta combinada + celda de resultado. Registra la fórmula para que
    `main.py` verifique que quedó con valor cacheado."""
    ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True, size=11)
    ws.cell(row=fila, column=1).alignment = Alignment(
        horizontal='left', vertical='center')
    if col_valor > 2:
        ref = f'A{fila}:{get_column_letter(col_valor - 1)}{fila}'
        if ref not in [str(r) for r in ws.merged_cells.ranges]:
            ws.merge_cells(ref)
    cel = ws.cell(row=fila, column=col_valor)
    cel.value = formula
    motor.calculada(cel, fmt or motor.FMT_ENT)
    cel.font = Font(bold=True, size=12)
    motor.reg(ws, cel.coordinate, formula)
    if rojo_si:
        motor.cf_formula(ws, cel.coordinate, rojo_si)
    return cel.coordinate


def _pie(ws, fila, ncols, extra=()):
    """Pie común: notas propias + conservación (§1.6) + marca."""
    for i, txt in enumerate(extra):
        motor.nota(ws, fila + i, txt, ncols=ncols)
    f = fila + len(extra)
    motor.nota(ws, f, motor.CONSERVACION, ncols=ncols)
    motor.nota(ws, f + 1, motor.MARCA, ncols=ncols)
    return f + 1


# ==========================================================================
# 03 — Plan maestro de Limpieza y Desinfección
# ==========================================================================
# DOM-05 (alta): el plan que se cuelga en la cocina y se entrega al inspector
# prescribía para los fregaderos «Desincrustante + lejía diluida». Un
# desincrustante es ácido y el hipoclorito con ácido libera cloro gaseoso: el
# documento normalizaba por escrito un accidente laboral clásico. Aquí el ácido
# y el cloro van en DOS pasadas, con el aclarado obligatorio en medio, y la
# advertencia viaja en la columna de método, que es la que se lee al trabajar.
CAB_03 = ['Zona / Elemento', 'Qué se limpia', 'Frecuencia',
          'Producto / Dilución', 'Tiempo de contacto',
          'Nº registro / FDS (S/N)', 'Método', 'Responsable', 'Verificación']
ANCHOS_03 = [30, 30, 16, 42, 16, 15, 48, 20, 22]
NC_03 = len(CAB_03)

# DOM-31/TEC-11: la lista de v1.1 tenía seis opciones y CINCO de las celdas
# precargadas contenían valores que no estaban en ella («2 veces/día», «Cada
# uso»…). Al desplegar, el usuario perdía la frecuencia correcta. La lista se
# amplía a lo que el plan usa de verdad.
FREC_03 = ('"Cada uso,Cada servicio,2 veces/día,Diaria,Semanal,Quincenal,'
           'Mensual,Trimestral,Semestral,Anual"')

# (zona, qué se limpia, frecuencia, producto, tiempo, FDS, método, resp, verif)
PLAN_03 = [
    ('sec', 'COCINA'),
    ('fila', ('Superficies de trabajo', 'Encimeras y tablas de corte',
              'Cada servicio', 'Desengrasante alimentario + desinfectante de '
              'superficies de uso alimentario', '5 min', 'S',
              'Retirar restos, pulverizar, dejar actuar el tiempo indicado, '
              'aclarar y secar con papel de un solo uso',
              'Cocinero de turno', 'Visual + test de superficie')),
    ('fila', ('Suelos de cocina', 'Suelo, rodapiés y rejillas', 'Diaria',
              'Desengrasante diluido 1:20', '10 min', 'S',
              'Barrer, fregar, aclarar y secar', 'Auxiliar de cocina',
              'Visual')),
    ('fila', ('Cámaras de refrigeración', 'Interior, estantes y juntas de '
              'puerta', 'Semanal', 'Desinfectante de uso alimentario',
              '5 min', 'S',
              'Vaciar, limpiar, aclarar y secar antes de reponer el producto',
              'Jefe de cocina', 'Visual + registro 04')),
    ('fila', ('Cámara de congelación', 'Interior, estantes y juntas',
              'Trimestral', 'Desinfectante de uso alimentario', '5 min', 'S',
              'Vaciar, desescarchar, limpiar y secar por completo',
              'Jefe de cocina', 'Visual + registro 04')),
    ('fila', ('Filtros de campana', 'Filtros metálicos', 'Semanal',
              'Desengrasante industrial', '20 min en inmersión', 'S',
              'Desmontar, sumergir, aclarar y secar antes de montar',
              'Auxiliar de cocina', 'Visual')),
    # DOM-31: en v1.1 esto era UNA fila con la frecuencia «Semanal filtros /
    # Mensual campana», que no estaba en el desplegable. Se parte en dos.
    ('fila', ('Campana y conductos de extracción', 'Superficie exterior y '
              'conductos', 'Mensual', 'Desengrasante industrial', '15 min',
              'S', 'Desengrasar la campana y la parte accesible del conducto; '
              'el conducto completo, una vez al año y por empresa autorizada',
              'Empresa externa', 'Certificado de la empresa')),
    ('fila', ('Fogones, plancha y horno', 'Quemadores, plancha e interior del '
              'horno', 'Diaria', 'Desengrasante de cocina', '10 min', 'S',
              'Rascar en frío, pulverizar, frotar y aclarar',
              'Cocinero de turno', 'Visual')),
    ('fila', ('Freidoras', 'Cuba, cestas y exterior', 'Semanal',
              'Desengrasante de cocina', '15 min', 'S',
              'Vaciar el aceite al bidón del gestor, rascar, desengrasar y '
              'aclarar', 'Cocinero de turno', 'Visual + registro 09')),
    # DOM-R2-12 (ronda 2): esta línea usa el desincrustante ácido, y el listado
    # de la pestaña «Productos químicos» entrega ese mismo producto con la FDS
    # marcada «N» y la nota «PENDIENTE: pedir la FDS al proveedor». Con «S»
    # aquí, el contador del pie daba 0 líneas pendientes y las dos pestañas del
    # mismo fichero se contradecían sobre el mismo producto. Con «N» el ejemplo
    # enseña el contador funcionando, que es para lo que está.
    ('fila', ('Fregaderos', 'Pilas, grifos y sifones', 'Diaria',
              'Desincrustante ácido (aclarar por completo) — desinfectar en '
              'una SEGUNDA pasada con hipoclorito 1:50', '5 min + 5 min', 'N',
              'Frotar con estropajo, ACLARAR POR COMPLETO y sólo entonces '
              'desinfectar. NUNCA mezclar ácido y lejía: la mezcla libera '
              'cloro gaseoso', 'Auxiliar de cocina', 'Visual')),
    ('fila', ('Lavamanos de accionamiento no manual', 'Pila, grifo de pedal o '
              'codo y dispensadores', 'Diaria', 'Desinfectante multiusos',
              '5 min', 'S',
              'Frotar, aclarar y reponer jabón, papel y papelera de pedal',
              'Auxiliar de cocina', 'Visual + dotación')),
    ('fila', ('Maquinaria: cortafiambres', 'Cuchilla, carro y protector',
              'Cada uso', 'Desinfectante de uso alimentario', '5 min', 'S',
              'Desconectar, desmontar las piezas, lavar, desinfectar y montar '
              'en seco', 'Cocinero de turno', 'Visual')),
    ('fila', ('Maquinaria: picadora y batidora', 'Cuerpo, cuchillas y '
              'accesorios', 'Cada uso', 'Desinfectante de uso alimentario',
              '5 min', 'S', 'Desmontar, lavar, desinfectar y secar al aire',
              'Cocinero de turno', 'Visual')),
    ('fila', ('Máquina de hielo', 'Depósito, pala y cazoleta', 'Mensual',
              'Desinfectante de uso alimentario apto para máquinas de hielo',
              '10 min', 'S',
              'Vaciar, desincrustar, desinfectar y aclarar con abundante agua; '
              'desechar la primera producción', 'Jefe de cocina',
              'Visual + registro 04')),
    ('fila', ('Paños, bayetas y estropajos', 'Textiles y estropajos de '
              'cocina', 'Cada servicio', 'Lavado a 60 °C o sustitución',
              'Ciclo completo', 'N/A',
              'Lavar a 60 °C como mínimo o desechar los de un solo uso; nunca '
              'reutilizar un paño de crudo en cocinado',
              'Auxiliar de cocina', 'Visual')),
    ('fila', ('Desagües y sumideros', 'Rejillas, arquetas y sifones',
              'Semanal', 'Desengrasante enzimático', '30 min', 'S',
              'Retirar la rejilla, extraer los residuos, aplicar y aclarar',
              'Auxiliar de cocina', 'Visual + olor')),
    ('fila', ('Cubos de residuos de cocina', 'Interior, exterior, tapa y '
              'pedal', 'Diaria', 'Hipoclorito diluido 1:50', '5 min', 'S',
              'Vaciar, lavar, desinfectar y secar', 'Auxiliar de cocina',
              'Visual + olor')),
    ('sec', 'SALA / COMEDOR'),
    ('fila', ('Mesas y sillas', 'Superficie, cantos y patas', 'Cada servicio',
              'Desinfectante multiusos de uso alimentario', '5 min', 'S',
              'Pulverizar, frotar y secar', 'Camarero', 'Visual')),
    ('fila', ('Barra y tiradores', 'Superficie, vitrinas y tiradores de '
              'cerveza', 'Cada servicio', 'Desinfectante multiusos', '5 min',
              'S', 'Pulverizar, frotar y secar; los tiradores, según el manual '
              'del fabricante', 'Barman', 'Visual')),
    ('fila', ('Suelos de sala', 'Suelo completo', 'Diaria',
              'Fregasuelos neutro', '10 min', 'S', 'Barrer, fregar y secar',
              'Personal de limpieza', 'Visual')),
    ('fila', ('Cristalería y vajilla', 'Vasos, platos y cubiertos', 'Cada uso',
              'Detergente de lavavajillas + abrillantador', 'Ciclo completo',
              'S', 'Lavavajillas con lavado a 60 °C y aclarado a 82 °C',
              'Personal de cocina', 'Temperatura del lavavajillas')),
    ('sec', 'BAÑOS'),
    ('fila', ('Inodoros y urinarios', 'Taza, exterior y cisterna',
              '2 veces/día', 'Limpiador de WC (aplicar solo) — desinfectar '
              'después con hipoclorito 1:50', '10 min', 'S',
              'Aplicar, frotar con escobilla, ACLARAR y sólo entonces '
              'desinfectar. Nunca mezclar dos productos en la taza',
              'Personal de limpieza', 'Visual + checklist')),
    ('fila', ('Lavabos, espejos y dispensadores', 'Pila, grifo, espejo, jabón '
              'y papel', '2 veces/día', 'Multiusos + limpiacristales', '5 min',
              'S', 'Frotar, aclarar, secar y reponer la dotación',
              'Personal de limpieza', 'Visual + dotación')),
    ('fila', ('Suelos de baños', 'Suelo y rejillas', '2 veces/día',
              'Hipoclorito diluido 1:50', '10 min', 'S',
              'Barrer, fregar y secar', 'Personal de limpieza',
              'Visual + olor')),
    ('sec', 'ALMACÉN'),
    ('fila', ('Estantes y baldas', 'Superficie de los estantes', 'Semanal',
              'Multiusos de uso alimentario', '5 min', 'S',
              'Retirar el producto, limpiar, secar y reponer respetando el '
              'FIFO', 'Almacenero', 'Visual')),
    ('fila', ('Suelos de almacén', 'Suelo completo', 'Semanal',
              'Fregasuelos neutro', '10 min', 'S', 'Barrer y fregar',
              'Personal de limpieza', 'Visual')),
    ('fila', ('Zona de recepción de mercancías', 'Suelo, mesa de descarga y '
              'báscula', 'Diaria', 'Desinfectante de uso alimentario', '5 min',
              'S', 'Retirar embalajes, limpiar y desinfectar tras cada '
              'descarga', 'Almacenero', 'Visual')),
    ('sec', 'VESTUARIOS'),
    ('fila', ('Taquillas y bancos', 'Superficie exterior', 'Semanal',
              'Multiusos', '5 min', 'S', 'Frotar y secar',
              'Personal de limpieza', 'Visual')),
    ('fila', ('Duchas y aseos del personal', 'Plato, paredes y grifería',
              'Diaria', 'Antical (aclarar por completo) — desinfectar después',
              '10 min + 5 min', 'S',
              'Frotar, aclarar y desinfectar en una segunda pasada',
              'Personal de limpieza', 'Visual')),
    ('sec', 'EXTERIOR'),
    ('fila', ('Terraza y entrada', 'Suelo, mesas, sillas y felpudos', 'Diaria',
              'Fregasuelos neutro', '10 min', 'S',
              'Barrer, fregar y retirar residuos', 'Personal de limpieza',
              'Visual')),
    ('fila', ('Contenedores y cubos exteriores', 'Interior, exterior y tapas',
              'Semanal', 'Hipoclorito diluido 1:50', '10 min', 'S',
              'Vaciar, lavar a presión, desinfectar y secar',
              'Personal de limpieza', 'Visual + olor')),
    ('fila', ('Zona de carga y descarga', 'Suelo, rampa y muelle', 'Semanal',
              'Desengrasante diluido 1:20', '10 min', 'S',
              'Barrer, fregar y retirar residuos', 'Almacenero', 'Visual')),
    ('fila', ('Cámara / cuarto de residuos', 'Suelo, paredes y contenedores',
              'Diaria', 'Hipoclorito diluido 1:50', '10 min', 'S',
              'Vaciar, fregar, desinfectar y ventilar',
              'Personal de limpieza', 'Visual + olor')),
    ('sec', 'ZONAS PROPIAS (añade aquí las de tu establecimiento)'),
    ('libre', None), ('libre', None), ('libre', None),
    ('libre', None), ('libre', None), ('libre', None),
]


def _plan_03():
    """Devuelve (filas_datos, filas_banda, ultima). Se calcula una sola vez y
    lo comparten `post`, los contadores y `CASOS_LIMITE`."""
    datos, bandas, fila = [], [], 5
    for tipo, _ in ((t, v) for t, v in PLAN_03):
        if tipo == 'sec':
            bandas.append(fila)
        else:
            datos.append(fila)
        fila += 1
    return datos, bandas, fila - 1


F_DATOS_03, F_BANDAS_03, F_ULT_03 = _plan_03()
F_CONT_03 = F_ULT_03 + 2                       # fila del contador de FDS


def _post_03(wb, fname, cambios):
    ws = wb['Plan Maestro L+D']
    _lienzo(ws, F_CONT_03 + 8, NC_03 + 2)
    _titulo(ws, 'Plan Maestro de Limpieza y Desinfección', NC_03,
            'Establecimiento: ________________________    '
            'Fecha de revisión: ___/___/______    '
            'Responsable del plan: ________________________')
    motor.cabecera(ws, 4, CAB_03, ANCHOS_03)

    fila = 5
    n_elem = 0
    for tipo, valor in PLAN_03:
        if tipo == 'sec':
            motor.banda(ws, fila, valor, NC_03)
            fila += 1
            continue
        for col in range(1, NC_03 + 1):
            cel = ws.cell(row=fila, column=col)
            motor.verde(cel, align='center' if col in (3, 5, 6) else 'left')
        if tipo == 'fila':
            for col, txt in enumerate(valor, start=1):
                ws.cell(row=fila, column=col).value = txt
            n_elem += 1
            ws.row_dimensions[fila].height = _alto(
                max(valor, key=len), 46, minimo=30)
        else:
            ws.row_dimensions[fila].height = 22
        fila += 1

    rangos_c = _rangos_filas('C', F_DATOS_03)
    rangos_f = _rangos_filas('F', F_DATOS_03)
    motor.dv_lista(ws, FREC_03, rangos_c, 'Frecuencia',
                   'Elige una frecuencia de la lista. Si la que necesitas no '
                   'está, parte la línea en dos: una frecuencia por fila es lo '
                   'que hace que el registro 04 pueda evidenciarla.')
    motor.dv_lista(ws, '"S,N,N/A"', rangos_f, 'Nº registro / FDS',
                   'Escribe S si el producto tiene nº de registro y su ficha '
                   'de datos de seguridad está accesible, N si falta alguna de '
                   'las dos, o N/A si en esa línea no se usa producto químico.')
    for r in rangos_f:
        motor.semaforo(ws, r, extra_ok=('S',), extra_rojo=('N',))

    f = (f'=COUNTIF(F{F_DATOS_03[0]}:F{F_ULT_03},"N")')
    _contador(ws, F_CONT_03, NC_03, 6,
              'Líneas con producto químico SIN nº de registro o SIN ficha de '
              'datos de seguridad accesible (N):', f,
              rojo_si=f'=F{F_CONT_03}>0')

    pie = _pie(ws, F_CONT_03 + 2, NC_03, extra=(
        'NUNCA mezclar productos: un desincrustante o un antical son ÁCIDOS y '
        'con hipoclorito (lejía) liberan cloro gaseoso. Donde hacen falta los '
        'dos, se aplican en pasadas separadas y con aclarado completo en medio.',
        'La columna «Tiempo de contacto» es el tiempo que el desinfectante '
        'tiene que quedarse mojando la superficie antes de aclarar. Sin ese '
        'dato la desinfección no es verificable, y es de lo primero que se '
        'pregunta en una inspección.',
        'Revisa la columna «Nº registro / FDS» con TUS productos reales: pon N '
        'en cuanto falte el número de registro o la ficha de datos de '
        'seguridad, y el contador de arriba te dirá cuántas te quedan. El '
        'detalle de cada producto va en la pestaña «Productos químicos».',
        'Cada línea de este plan se evidencia en el registro 04 (Limpieza '
        'diaria): las de frecuencia diaria o por servicio, en la rejilla de '
        'días; las semanales y mensuales, en el bloque «TAREAS SEMANALES».'))
    motor.IMPRESION[(fname, ws.title)] = (4, True)
    cambios.append(f'03: plan reconstruido con {n_elem} elementos en 6 zonas '
                   '(incluida EXTERIOR), fregaderos sin ácido+lejía, columnas '
                   'de tiempo de contacto y FDS y DV de frecuencia ampliada '
                   '(DOM-05/DOM-11/DOM-12/DOM-31/TEC-11/TEC-12/COM-06)')
    _hoja_quimicos(wb, fname, cambios)
    _instrucciones_03(wb, cambios, n_elem, pie)


CAB_03Q = ['Producto (nombre comercial)', 'Para qué se usa',
           'Nº de registro (HA / plaguicidas / sanitario)', 'Dosis / dilución',
           'Tiempo de contacto', 'FDS disponible (S/N)',
           'Dónde está la ficha de datos de seguridad', 'Observaciones']
ANCHOS_03Q = [30, 28, 34, 20, 18, 16, 34, 30]
HOJA_QUIM = 'Productos químicos'
F0_03Q, F1_03Q = 5, 28


def _hoja_quimicos(wb, fname, cambios):
    """DOM-12: el inspector pide que el producto sea de uso alimentario, con su
    número de registro, y que las fichas de datos de seguridad estén
    accesibles. El plan v1.1 sólo tenía descripciones genéricas."""
    ws = motor.hoja(wb, HOJA_QUIM)
    _lienzo(ws, F1_03Q + 8, len(CAB_03Q) + 2)
    _titulo(ws, 'Listado de Productos Químicos de Limpieza y Desinfección',
            len(CAB_03Q),
            'Anota aquí TODOS los productos que aparecen en el plan maestro. '
            'Las fichas de datos de seguridad (FDS) las entrega el proveedor y '
            'tienen que estar accesibles para el personal y para la '
            'inspección.')
    motor.cabecera(ws, 4, CAB_03Q, ANCHOS_03Q)
    for col in range(1, len(CAB_03Q) + 1):
        motor.verde(ws.cell(row=F0_03Q, column=col),
                    align='center' if col in (5, 6) else 'left')
    motor.replicar_filas(ws, F0_03Q, F0_03Q, F1_03Q, ncols=len(CAB_03Q),
                         alto=26)
    motor.dv_lista(ws, '"S,N"',
                   [motor.rango('F', F0_03Q, F1_03Q)], 'FDS disponible',
                   'Escribe S si tienes la ficha de datos de seguridad '
                   'accesible o N si te falta.')
    motor.semaforo(ws, motor.rango('F', F0_03Q, F1_03Q),
                   extra_ok=('S',), extra_rojo=('N',))
    motor.sembrar(ws, 5, {
        'A': 'Desengrasante alimentario', 'B': 'Superficies y suelos de cocina',
        'C': 'Registro HA nº 00-00-00000', 'D': '1:20', 'E': '5 min', 'F': 'S',
        'G': 'Carpeta APPCC, pestaña «Productos químicos»'}, marca_col='H')
    motor.sembrar(ws, 6, {
        'A': 'Desinfectante de superficies', 'B': 'Desinfección de superficies '
        'en contacto con alimentos', 'C': 'Registro HA nº 00-00-00000',
        'D': 'Listo al uso', 'E': '5 min', 'F': 'S',
        'G': 'Carpeta APPCC, pestaña «Productos químicos»'}, marca_col='H')
    motor.sembrar(ws, 7, {
        'A': 'Hipoclorito sódico (lejía alimentaria)',
        'B': 'Cubos, suelos de baño y contenedores', 'D': '1:50',
        'C': 'Registro sanitario nº 00-00000/X', 'E': '5 min', 'F': 'S',
        'G': 'Cuarto de limpieza, junto al armario cerrado'}, marca_col='H')
    motor.sembrar(ws, 8, {
        'A': 'Desincrustante ácido', 'B': 'Fregaderos y grifería',
        'C': 'Registro HA nº 00-00-00000', 'D': 'Listo al uso', 'E': '5 min',
        'F': 'N', 'G': 'PENDIENTE: pedir la FDS al proveedor'}, marca_col='H')
    _pie(ws, F1_03Q + 2, len(CAB_03Q), extra=(
        'Los productos de limpieza se guardan en un armario cerrado y NUNCA '
        'junto a los alimentos, en su envase original y con su etiqueta. '
        'Trasvasarlos a una botella sin etiquetar es uno de los incumplimientos '
        'que más se sancionan.',))
    motor.IMPRESION[(fname, HOJA_QUIM)] = (4, True)
    cambios.append('03: pestaña «Productos químicos» con nº de registro, '
                   'dosis, tiempo de contacto y ubicación de la FDS (DOM-12)')


def _instrucciones_03(wb, cambios, n_elem, ultima):
    motor.escribir_instrucciones(wb, 'Plan de Limpieza y Desinfección (L+D)', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'La pestaña «Plan Maestro L+D» es el documento oficial: define '
              'qué se limpia, cada cuánto, con qué producto, durante cuánto '
              'tiempo, cómo y quién lo hace. Se imprime y se cuelga en la '
              'cocina, y se entrega tal cual en una inspección.'),
        ('b', f'Trae {n_elem} elementos repartidos en seis zonas (cocina, sala, '
              'baños, almacén, vestuarios y exterior) más filas libres al final '
              'para las tuyas. Todas las celdas verdes son editables.'),
        ('b', 'La columna «Frecuencia» es un desplegable y ahora sí recoge las '
              'frecuencias que el plan usa de verdad: cada uso, cada servicio, '
              '2 veces/día, diaria, semanal, quincenal, mensual, trimestral, '
              'semestral y anual.'),
        ('b', 'La pestaña «Productos químicos» lista cada producto con su '
              'número de registro, su dosis, su tiempo de contacto y dónde '
              'está su ficha de datos de seguridad.'),
        ('h', 'Lo que NUNCA se mezcla'),
        ('p', 'Un desincrustante, un antical o un limpiador de WC son ÁCIDOS. '
              'Mezclados con hipoclorito (lejía) liberan cloro gaseoso, que es '
              'tóxico. En este plan, donde hacen falta los dos, van en dos '
              'pasadas separadas con aclarado completo en medio, y así está '
              'escrito en la columna «Método».'),
        ('h', 'Tiempo de contacto y ficha de datos de seguridad'),
        ('b', 'El tiempo de contacto es el que el desinfectante tiene que '
              'quedarse mojando la superficie antes de aclarar. Sin ese dato '
              'la desinfección no es verificable.'),
        ('b', 'El inspector comprueba que el producto sea de uso alimentario, '
              'que tenga número de registro y que su ficha de datos de '
              'seguridad esté accesible. Marca N en la columna «Nº registro / '
              'FDS» mientras te falte alguna: el contador del final de la hoja '
              'te dice cuántas llevas pendientes.'),
        ('h', 'Cómo se evidencia este plan'),
        ('b', 'Un plan sin registro no acredita nada. Cada línea se evidencia '
              'en el registro 04 (Limpieza diaria): las frecuencias diarias y '
              'por servicio en la rejilla de días, y las semanales y mensuales '
              'en el bloque «TAREAS SEMANALES» de esa misma hoja.'),
        ('b', 'Revisa el plan al menos una vez al año, y siempre que cambies '
              'de producto, de proveedor de limpieza o de instalación.'),
        ('b', motor.CONSERVACION),
    ], cambios)


# ==========================================================================
# 04 — Registro de limpieza diaria
# ==========================================================================
# DOM-13/TEC-27: el registro no podía evidenciar el plan que acompaña. El plan
# exige «Cada servicio» para superficies, mesas y barra, y «2 veces/día» para
# los baños; el registro daba UNA casilla por tarea y día. Y las tareas
# semanales del plan (cámaras, almacén, vestuarios, filtros) no tenían dónde
# anotarse. Ahora hay dos columnas por día (M/T) y un bloque semanal con fecha
# y firma.
DIAS_04 = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
NC_04 = 1 + 2 * len(DIAS_04)                   # A + 14 columnas M/T
MARCAS_04 = '"✓,✗,N/A"'

TAREAS_04 = [
    ('sec', 'COCINA'),
    ('t', 'Superficies de trabajo y tablas desinfectadas'),
    ('t', 'Fogones, plancha y horno limpiados'),
    ('t', 'Fregaderos limpiados (ácido y desinfectante en pasadas separadas)'),
    ('t', 'Lavamanos: jabón, papel y papelera de pedal repuestos'),
    ('t', 'Cortafiambres, picadora y batidora desmontados y desinfectados'),
    ('t', 'Cubos de residuos vaciados, lavados y desinfectados'),
    ('t', 'Suelos barridos y fregados'),
    ('t', 'Cámaras: orden, limpieza exterior y juntas de puerta'),
    ('t', 'Paños y bayetas cambiados o lavados a 60 °C'),
    ('sec', 'SALA / COMEDOR'),
    ('t', 'Mesas y sillas limpiadas'),
    ('t', 'Barra y tiradores desinfectados'),
    ('t', 'Suelos barridos y fregados'),
    ('t', 'Cristalería y vajilla repasadas'),
    ('sec', 'BAÑOS'),
    ('t', 'Inodoros y urinarios limpiados'),
    ('t', 'Lavabos y espejos limpiados'),
    ('t', 'Suelos fregados'),
    ('t', 'Jabón, papel y papelera repuestos'),
    ('sec', 'EXTERIOR'),
    ('t', 'Terraza, entrada y felpudos barridos'),
    ('t', 'Cubos exteriores y cuarto de residuos'),
    # DOM-R2-13 (ronda 2): el plan 03 promete DOS veces que «cada línea de este
    # plan se evidencia en el registro 04», y cuatro de sus 32 líneas no tenían
    # dónde. Estas dos son diarias en el plan y no tenían casilla diaria aquí
    # (no existe bloque ALMACÉN, y las duchas sólo aparecían en el bloque
    # semanal, mezcladas con taquillas y bancos).
    ('t', 'Zona de recepción de mercancías: suelo y mesa de descarga'),
    ('t', 'Duchas y aseos del personal'),
]

SEMANALES_04 = [
    'Cámaras de refrigeración: interior, estantes y juntas',
    'Filtros de campana desmontados y desengrasados',
    'Almacén: estantes, baldas y suelos',
    'Vestuarios: taquillas, bancos y duchas',
    'Desagües y sumideros de cocina',
    'Freidoras: vaciado, rascado y limpieza de cuba',
    'Contenedores y zona de carga exteriores',
    'Máquina de hielo (mensual: marca la semana que toque)',
    # DOM-R2-13: las otras dos líneas huérfanas del plan. La campana es mensual
    # (el conducto, anual y por empresa autorizada) y la cámara de congelación
    # trimestral: el bloque se anunciaba como «semanal, quincenal y mensual» y
    # no tenía fila para ninguna de las dos.
    'Campana y conductos de extracción (mensual; el conducto, anual y por '
    'empresa autorizada)',
    'Cámara de congelación: interior, estantes y juntas (trimestral)',
]

F0_04 = 6                                       # primera fila de la rejilla
F1_04 = F0_04 + len(TAREAS_04) - 1              # última (28)
FV_04 = F1_04 + 2                               # banda VERIFICACIÓN (30)
FS_04 = FV_04 + 4                               # banda TAREAS SEMANALES (34)
FSH_04 = FS_04 + 1                              # cabecera del bloque semanal
FS0_04, FS1_04 = FSH_04 + 1, FSH_04 + len(SEMANALES_04)
F_CONT_04 = FS1_04 + 2

# Bloques de la rejilla que llevan DV y semáforo (las bandas de sección van
# combinadas de A a O y quedan fuera).
FILAS_TAREA_04 = [F0_04 + i for i, (t, _) in enumerate(TAREAS_04) if t == 't']


def _cab_celda(ws, fila, col, texto):
    cel = ws.cell(row=fila, column=col, value=texto)
    cel.font = Font(bold=True, color='FFFFFF', size=10)
    cel.fill = PatternFill('solid', fgColor=motor.CAB)
    cel.alignment = Alignment(horizontal='center', vertical='center',
                              wrap_text=True)
    cel.border = motor.BORDE
    return cel


def _post_04(wb, fname, cambios):
    ws = wb['Limpieza Diaria']
    _lienzo(ws, F_CONT_04 + 10, NC_04 + 2)
    _titulo(ws, 'Registro de Limpieza Diaria', NC_04,
            'Semana del: ___/___/______ al ___/___/______        '
            'Establecimiento: ________________________')

    ws.column_dimensions['A'].width = 44
    for c in range(2, NC_04 + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6.5
    _cab_celda(ws, 4, 1, 'Tarea de Limpieza')
    ws.merge_cells('A4:A5')
    for i, dia in enumerate(DIAS_04):
        col = 2 + 2 * i
        _cab_celda(ws, 4, col, dia)
        _cab_celda(ws, 4, col + 1, None)
        ws.merge_cells(start_row=4, start_column=col,
                       end_row=4, end_column=col + 1)
        _cab_celda(ws, 5, col, 'M')
        _cab_celda(ws, 5, col + 1, 'T')
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 16

    fila = F0_04
    for tipo, texto in TAREAS_04:
        if tipo == 'sec':
            motor.banda(ws, fila, texto, NC_04)
            ws.row_dimensions[fila].height = 18
            fila += 1
            continue
        cel = ws.cell(row=fila, column=1, value=texto)
        motor.calculada(cel)
        cel.alignment = Alignment(horizontal='left', vertical='center',
                                  wrap_text=True)
        for c in range(2, NC_04 + 1):
            motor.verde(ws.cell(row=fila, column=c), '@')
        ws.row_dimensions[fila].height = _alto(texto, 42, minimo=20)
        fila += 1

    # TEC-22: la hoja pedía marcar con ✓ un carácter que no está en un teclado
    # español, sin desplegable ni casilla. El desplegable resuelve las dos
    # cosas: se marca con el ratón y el contador de abajo puede comparar por
    # igualdad exacta. El rango cubre B6:O28 (y por tanto B6:H28).
    rangos_marca = []
    for c in range(2, NC_04 + 1):
        col = get_column_letter(c)
        rangos_marca += _rangos_filas(col, FILAS_TAREA_04)
    motor.dv_lista(ws, MARCAS_04, rangos_marca, 'Tarea realizada',
                   'Marca ✓ (hecha), ✗ (no hecha) o N/A (ese turno no aplica). '
                   'Si escribes otra cosa, el contador de tareas pendientes no '
                   'la reconoce.')
    for c in range(2, NC_04 + 1):
        motor.semaforo(ws, motor.rango(get_column_letter(c), F0_04, F1_04))

    # §1.5 — ejemplo: la columna del lunes de las tres primeras tareas.
    for f in FILAS_TAREA_04[:3]:
        ws.cell(row=f, column=2).value = '✓'
        ws.cell(row=f, column=3).value = '✓'

    motor.banda(ws, FV_04, 'VERIFICACIÓN DEL RESPONSABLE', NC_04)
    for i, etiqueta in enumerate(('Verificado por (iniciales)',
                                  'Firma del responsable')):
        f = FV_04 + 1 + i
        cel = ws.cell(row=f, column=1, value=etiqueta)
        motor.calculada(cel)
        cel.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(2, NC_04 + 1):
            motor.verde(ws.cell(row=f, column=c), '@')
        ws.row_dimensions[f].height = 22

    # DOM-13: el bloque que faltaba. Las frecuencias semanales y mensuales del
    # plan 03 no tenían dónde anotarse, así que quedaban sin registro alguno.
    motor.banda(ws, FS_04, 'TAREAS SEMANALES, MENSUALES Y PERIÓDICAS '
                           '(registro 03: todo lo que no es diario)', NC_04)
    _cab_celda(ws, FSH_04, 1, 'Tarea')
    for txt, ini, fin in (('Fecha', 2, 4), ('Hecho', 5, 6),
                          ('Responsable', 7, 10), ('Firma', 11, NC_04)):
        _cab_celda(ws, FSH_04, ini, txt)
        for c in range(ini + 1, fin + 1):
            _cab_celda(ws, FSH_04, c, None)
        ws.merge_cells(start_row=FSH_04, start_column=ini,
                       end_row=FSH_04, end_column=fin)
    ws.row_dimensions[FSH_04].height = 20

    for i, texto in enumerate(SEMANALES_04):
        f = FS0_04 + i
        cel = ws.cell(row=f, column=1, value=texto)
        motor.calculada(cel)
        cel.alignment = Alignment(horizontal='left', vertical='center',
                                  wrap_text=True)
        for ini, fin in ((2, 4), (5, 6), (7, 10), (11, NC_04)):
            for c in range(ini, fin + 1):
                motor.verde(ws.cell(row=f, column=c), '@')
            ws.merge_cells(start_row=f, start_column=ini,
                           end_row=f, end_column=fin)
        ws.row_dimensions[f].height = 22
    motor.dv_lista(ws, MARCAS_04, [motor.rango('E', FS0_04, FS1_04)],
                   'Tarea realizada',
                   'Marca ✓ (hecha), ✗ (no hecha) o N/A.')
    motor.semaforo(ws, motor.rango('E', FS0_04, FS1_04))
    # DOM-R2-20: el aviso de ejemplo del pie sólo cubría la rejilla diaria, y
    # esta línea del bloque semanal iba sembrada con fecha, marca y responsable
    # sin distintivo: quien borrase lo que la nota decía imprimía y archivaba
    # una limpieza de cámaras que no hizo nadie. Ahora lo dice la propia fila.
    motor.sembrar(ws, FS0_04, {'B': '07/09/2026', 'E': '✓',
                               'G': 'A.R. (ejemplo)'})

    suma = '+'.join(
        f'COUNTIF({get_column_letter(c)}{F0_04}:{get_column_letter(c)}{F1_04},'
        f'"✗")' for c in range(2, NC_04 + 1))
    f_cont = f'={suma}+COUNTIF(E{FS0_04}:E{FS1_04},"✗")'
    _contador(ws, F_CONT_04, NC_04, 2,
              'Tareas marcadas como NO realizadas (✗) esta semana:', f_cont,
              rojo_si=f'=B{F_CONT_04}>0')

    _pie(ws, F_CONT_04 + 2, NC_04, extra=(
        'Cada día tiene dos casillas: M (mañana) y T (tarde). Las tareas que '
        'el plan 03 marca como «Cada servicio» o «2 veces/día» se marcan en '
        'las dos; las de una sola vez al día, en el turno en que se hicieron, '
        'y la otra casilla se deja en N/A.',
        'Las marcas del lunes de las tres primeras tareas y la primera línea '
        'del bloque semanal (marcada «(ejemplo)» en el responsable) son '
        'EJEMPLOS: bórralas antes de imprimir la semana.',
        'Si aparece ✗: anota la incidencia en el registro 11 (Acciones '
        'Correctivas) y repite la tarea antes del siguiente servicio.'))
    motor.IMPRESION[(fname, ws.title)] = ('4:5', True, True)
    cambios.append('04: 14 columnas M/T por día, bloque de tareas semanales '
                   'con fecha y firma, DV ✓/✗/N/A sobre todas las casillas y '
                   'contador de pendientes (DOM-13/TEC-22/TEC-27)')

    motor.escribir_instrucciones(wb, 'Registro de Limpieza Diaria', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Imprime una hoja por semana y cuélgala en la cocina, o rellénala '
              'en pantalla: las casillas son desplegables con ✓, ✗ y N/A.'),
        ('b', 'Cada día tiene DOS casillas, M (mañana) y T (tarde), porque el '
              'plan maestro (registro 03) exige «Cada servicio» en superficies, '
              'mesas y barra y «2 veces/día» en los baños. Con una sola casilla '
              'este registro no podía evidenciar su propio plan.'),
        ('b', 'Las tareas de una sola vez al día se marcan en el turno en que '
              'se hicieron y la otra casilla se deja en N/A.'),
        ('b', 'El bloque «TAREAS SEMANALES, MENSUALES Y PERIÓDICAS» recoge '
              'todo lo que no es diario: cámaras de refrigeración y de '
              'congelación, filtros y conductos de campana, almacén, '
              'vestuarios, desagües, freidoras, contenedores y máquina de '
              'hielo. Lleva fecha, responsable y firma. Cada línea del plan '
              'maestro (registro 03) tiene aquí dónde evidenciarse: las '
              'diarias en la rejilla de arriba y el resto en este bloque.'),
        ('b', 'El responsable verifica y firma en el bloque «VERIFICACIÓN». '
              'Una hoja sin firmar cuenta como no hecha.'),
        ('h', 'Qué hacer si algo no se ha hecho'),
        ('b', 'Marca ✗ y no lo disimules: el contador del final de la hoja los '
              'suma y se pone en rojo. Una hoja con un ✗ y su acción correctiva '
              'anotada acredita que el sistema funciona; una hoja con todo ✓ '
              'todas las semanas del año no se la cree nadie.'),
        ('b', 'Anota la incidencia en el registro 11 (Acciones Correctivas) y '
              'repite la tarea antes del siguiente servicio.'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Una hoja por semana. El área de impresión llega hasta el pie, '
              'así que lo que archivas incluye la nota de conservación.'),
        ('b', motor.CONSERVACION),
    ], cambios)


# ==========================================================================
# 07 — Control de plagas (DDD)
# ==========================================================================
# DOM-16: faltaban los tres datos que el inspector comprueba —el nº de
# inscripción de la empresa en el ROESB, el nº de registro del biocida aplicado
# y el plazo de seguridad de reentrada— y la landing prometía un «espacio para
# plano de cebos» que no existía: sólo había una frase pidiéndolo.
# TEC-20: la hoja se encabezaba «Año» y daba 20 filas, cuando sus propias
# Instrucciones prescriben ~70 actuaciones al año. Se queda sin sitio en marzo.
CAB_07 = ['Fecha', 'Tipo de actuación', 'Empresa / Técnico',
          'Producto utilizado', 'Nº registro del biocida',
          'Plazo de seguridad (h)', 'Zonas tratadas', 'Resultado / hallazgos',
          'Próxima revisión', 'Nº certificado']
ANCHOS_07 = [14, 22, 26, 30, 24, 18, 26, 34, 16, 18]
NC_07 = len(CAB_07)
F0_07, F1_07 = 5, 84                            # 80 filas (TEC-20)
F_CONT_07 = F1_07 + 2

HOJA_CEBOS = 'Plano de cebos'
CAB_07C = ['Nº estación', 'Ubicación', 'Tipo de dispositivo',
           'Fecha de instalación', 'Fecha última revisión', 'Estado / consumo',
           'Responsable', 'Observaciones']
ANCHOS_07C = [12, 36, 30, 18, 20, 22, 20, 30]
NC_07C = len(CAB_07C)
F0_07C, F1_07C = 5, 44                          # 40 estaciones
F_CONT_07C = F1_07C + 2

TIPO_07 = ('"Desinsectación,Desratización,Desinfección,Revisión de cebos,'
           'Inspección visual,Otro"')
TIPO_CEBO = ('"Cebadero de roedores,Trampa de captura,Lámpara insectocutora,'
             'Trampa de feromonas,Otro"')
ESTADO_CEBO = ('"Sin consumo,Consumo parcial,Consumo total,Captura,Dañado,'
               'Ausente"')

CEBOS_SEMBRADOS = [
    ('Puerta de servicio, exterior derecha', 'Cebadero de roedores'),
    ('Puerta de servicio, exterior izquierda', 'Cebadero de roedores'),
    ('Cuarto de residuos, junto al contenedor', 'Cebadero de roedores'),
    ('Almacén de secos, rodapié norte', 'Cebadero de roedores'),
    ('Zona de carga y descarga', 'Cebadero de roedores'),
    ('Cocina, entrada de personal', 'Lámpara insectocutora'),
    ('Zona de emplatado', 'Lámpara insectocutora'),
    ('Almacén de secos, estante de harinas', 'Trampa de feromonas'),
]


def _post_07(wb, fname, cambios):
    ws = wb['Control Plagas DDD']
    _lienzo(ws, F_CONT_07 + 10, NC_07 + 2)
    _titulo(ws, 'Registro de Control de Plagas (DDD)', NC_07,
            'Año: ______    Empresa contratada: ________________________    '
            'Nº ROESB: ______________    Nº de contrato: ____________',
            'El Nº ROESB es la inscripción de la empresa en el Registro Oficial '
            'de Establecimientos y Servicios Biocidas. Sin él, el servicio no '
            'está habilitado para aplicar biocidas de uso profesional.')
    motor.cabecera(ws, 4, CAB_07, ANCHOS_07)

    for col in range(1, NC_07 + 1):
        cel = ws.cell(row=F0_07, column=col)
        if col == 6:
            motor.verde(cel, motor.FMT_ENT)
        elif col in (1, 9):
            motor.verde(cel, '@')
        elif col in (2, 5, 10):
            motor.verde(cel, align='center')
        else:
            motor.verde(cel, align='left')
    motor.replicar_filas(ws, F0_07, F0_07, F1_07, ncols=NC_07, alto=22)

    motor.dv_lista(ws, TIPO_07, [motor.rango('B', F0_07, F1_07)],
                   'Tipo de actuación', 'Elige una actuación de la lista.')
    motor.dv_decimal(ws, [motor.rango('F', F0_07, F1_07)],
                     'Plazo de seguridad',
                     'Introduce el plazo de seguridad de reentrada en horas, '
                     'entre 0 y 168. Lo indica la ficha del biocida; hasta que '
                     'pase, no se puede manipular alimento en la zona tratada.',
                     minimo=0, maximo=168)

    # DOM-R2-14 (ronda 2): las dos actuaciones DDD iban sembradas SIN la marca
    # «(ejemplo)» —la llevaba sólo la tercera fila, la inspección visual—, y son
    # justo las dos peligrosas: documentan la intervención de una empresa
    # externa con su nº de certificado. Un cliente que archivase la hoja tenía
    # en su carpeta APPCC dos certificados de desratización y desinsectación
    # inventados, a nombre de una empresa que no ha contratado.
    motor.sembrar(ws, 5, {
        'A': '15/01/2026', 'B': 'Desratización',
        'C': 'Control de Plagas del Norte S.L. — téc. J.M.',
        'D': 'Bromadiolona 0,005 % en bloque parafinado',
        'E': 'ES/BIO-00-00000', 'F': 0,
        'G': 'Perímetro exterior, cuarto de residuos y almacén',
        'H': 'Sin consumo en las 5 estaciones; sin indicios',
        'I': '15/02/2026', 'J': 'CERT-2026-0115'}, marca_col='H')
    motor.sembrar(ws, 6, {
        'A': '15/01/2026', 'B': 'Desinsectación',
        'C': 'Control de Plagas del Norte S.L. — téc. J.M.',
        'D': 'Cipermetrina 10 % microencapsulada',
        'E': 'ES/BIO-00-00000', 'F': 6,
        'G': 'Desagües de cocina, cuarto de residuos y falsos techos',
        'H': 'Tratamiento preventivo trimestral; sin hallazgos',
        'I': '15/04/2026', 'J': 'CERT-2026-0116'}, marca_col='H')
    motor.sembrar(ws, 7, {
        'A': '22/01/2026', 'B': 'Inspección visual',
        'C': 'Personal propio — jefe de cocina', 'D': 'No aplica',
        'E': 'No aplica', 'F': 0,
        'G': 'Almacén de secos y rodapiés de cocina',
        'H': 'Un saco de harina roído: retirado, lote desechado y avisada la '
             'empresa · incidencia INC-004',
        'I': '29/01/2026', 'J': 'Interno'}, marca_col='H')

    _pie(ws, F_CONT_07, NC_07, extra=(
        'Adjunta el certificado de cada actuación y el plano de cebos '
        'actualizado (pestaña «Plano de cebos»). El inspector pide los tres: '
        'contrato con el Nº ROESB, certificados y plano.',
        'El «Plazo de seguridad» es el tiempo que debe pasar entre la '
        'aplicación del biocida y la vuelta a manipular alimentos en esa zona. '
        'Lo fija la ficha del producto y se anota en horas.',
        'Si aparecen indicios (excrementos, roeduras, insectos vivos), no '
        'esperes a la revisión programada: avisa a la empresa el mismo día y '
        'abre la incidencia en el registro 11 (Acciones Correctivas).',
        'Las tres primeras filas son EJEMPLOS y lo dicen en «Resultado / '
        'observaciones»: bórralas antes de imprimir la hoja. Los nombres de '
        'empresa, los nº de certificado y los nº de registro de biocida son '
        'inventados.'))
    motor.IMPRESION[(fname, ws.title)] = (4, True, 'C5')  # TEC-08

    _hoja_cebos(wb, fname, cambios)
    cambios.append('07: Nº ROESB en cabecera, columnas de nº de registro del '
                   f'biocida y plazo de seguridad, {F1_07 - F0_07 + 1} filas y '
                   'pestaña «Plano de cebos» (DOM-16/TEC-20)')

    motor.escribir_instrucciones(wb, 'Registro de Control de Plagas (DDD)', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Registra TODAS las actuaciones: desinsectación, desratización, '
              'desinfección, revisiones de cebos e inspecciones visuales '
              'propias. La hoja trae 80 filas, que es lo que da un año real.'),
        ('b', 'Anota en la cabecera el Nº ROESB de la empresa contratada. Es '
              'su inscripción en el Registro Oficial de Establecimientos y '
              'Servicios Biocidas y es lo primero que se comprueba: sin él, la '
              'empresa no está habilitada para aplicar biocidas profesionales.'),
        ('b', 'De cada aplicación anota el nº de registro del biocida y su '
              'plazo de seguridad en horas. Hasta que ese plazo pase no se '
              'puede manipular alimento en la zona tratada.'),
        ('b', 'La pestaña «Plano de cebos» es el croquis numerado de las '
              'estaciones: dónde está cada una, de qué tipo es, cuándo se '
              'revisó y qué se encontró. Imprímela y dibuja encima la planta '
              'del local con los números.'),
        ('h', 'Frecuencia recomendada'),
        ('b', 'Desratización: revisión mensual de los cebaderos.'),
        ('b', 'Desinsectación: trimestral, o antes si hay indicios.'),
        ('b', 'Desinfección general: semestral.'),
        ('b', 'Inspección visual propia: semanal, buscando excrementos, '
              'roeduras, insectos vivos y puntos de entrada.'),
        ('h', 'Qué guardar'),
        ('b', 'Contrato en vigor con el Nº ROESB, certificado de cada '
              'actuación, fichas de datos de seguridad de los biocidas y el '
              'plano de cebos actualizado.'),
        ('b', motor.CONSERVACION),
    ], cambios)


def _hoja_cebos(wb, fname, cambios):
    ws = motor.hoja(wb, HOJA_CEBOS)
    _lienzo(ws, F_CONT_07C + 8, NC_07C + 2)
    _titulo(ws, 'Plano de Cebos y Dispositivos de Control de Plagas', NC_07C,
            'Numera cada estación en el plano de tu local y usa el mismo número '
            'en esta tabla. Imprime la hoja y dibuja la planta al dorso o '
            'grapa el croquis de la empresa de control de plagas.')
    motor.cabecera(ws, 4, CAB_07C, ANCHOS_07C)

    for col in range(1, NC_07C + 1):
        cel = ws.cell(row=F0_07C, column=col)
        if col == 1:
            motor.calculada(cel, motor.FMT_ENT)
        elif col in (2, 8):
            motor.verde(cel, align='left')
        elif col in (4, 5):
            motor.verde(cel, '@')
        else:
            motor.verde(cel, align='center')
    motor.replicar_filas(ws, F0_07C, F0_07C, F1_07C, ncols=NC_07C, alto=22)
    for i in range(F0_07C, F1_07C + 1):
        ws.cell(row=i, column=1).value = i - F0_07C + 1

    motor.dv_lista(ws, TIPO_CEBO, [motor.rango('C', F0_07C, F1_07C)],
                   'Tipo de dispositivo', 'Elige un tipo de la lista.')
    motor.dv_lista(ws, ESTADO_CEBO, [motor.rango('F', F0_07C, F1_07C)],
                   'Estado / consumo',
                   'Elige un estado de la lista. El contador del final compara '
                   'por igualdad exacta, así que un texto libre no suma.')
    motor.semaforo(ws, motor.rango('F', F0_07C, F1_07C),
                   extra_ok=('Sin consumo',), extra_ambar=('Consumo parcial',),
                   extra_rojo=('Consumo total', 'Captura', 'Dañado', 'Ausente'))

    for i, (ubic, tipo) in enumerate(CEBOS_SEMBRADOS):
        motor.sembrar(ws, F0_07C + i, {
            'B': ubic, 'C': tipo, 'D': '15/01/2026', 'E': '15/01/2026',
            'F': 'Sin consumo', 'G': 'J.M. (empresa DDD)'}, marca_col='H')

    f = (f'=COUNTIF(F{F0_07C}:F{F1_07C},"Consumo parcial")'
         f'+COUNTIF(F{F0_07C}:F{F1_07C},"Consumo total")'
         f'+COUNTIF(F{F0_07C}:F{F1_07C},"Captura")')
    _contador(ws, F_CONT_07C, NC_07C, 6,
              'Estaciones con actividad (consumo o captura) en la última '
              'revisión:', f, rojo_si=f'=F{F_CONT_07C}>0')

    _pie(ws, F_CONT_07C + 2, NC_07C, extra=(
        'Si el contador es mayor que cero hay actividad: avisa a la empresa de '
        'control de plagas el mismo día, revisa los puntos de entrada y abre '
        'la incidencia en el registro 11 (Acciones Correctivas).',
        'Una estación «Dañada» o «Ausente» también es un hallazgo: hay que '
        'reponerla y dejarlo anotado.'))
    motor.IMPRESION[(fname, HOJA_CEBOS)] = (4, True)


# ==========================================================================
# 11 — Registro de acciones correctivas
# ==========================================================================
# DOM-24: faltaban las dos columnas que el inspector busca primero —qué
# producto o lote quedó afectado y qué se hizo con él— y sin la disposición del
# producto no conforme la acción correctiva está incompleta (principio 5 del
# APPCC). Tampoco había acción preventiva ni referencia al registro donde se
# detectó la desviación, que es lo que engancha esta hoja con el resto del pack.
CAB_11 = ['Nº', 'Fecha', 'Tipo de incidencia', 'Registro de origen',
          'Descripción del problema', 'Producto / lote afectado',
          'Destino del producto no conforme', 'Causa probable',
          'Acción correctiva adoptada', 'Acción preventiva', 'Responsable',
          'Fecha de cierre', 'Verificación']
ANCHOS_11 = [12, 13, 20, 24, 34, 24, 26, 26, 34, 32, 16, 14, 20]
NC_11 = len(CAB_11)
F0_11, F1_11 = 5, 44                            # 40 incidencias
F_CONT_11 = F1_11 + 2

TIPO_11 = ('"Temperatura,Producto rechazado,Limpieza,Plagas,'
           'Reclamación de cliente,Alérgeno,Contaminación,Caducidad,'
           'Trazabilidad,Formación,Otro"')
ORIGEN_11 = ('"01 Temperaturas diario,02 Recepción temperaturas,'
             '04 Limpieza diaria,05 Recepción mercancías,06 Trazabilidad,'
             '07 Plagas,08 Alérgenos,09 Aceite,10 Agua,16 Cocción,'
             '17 Enfriamiento,18 Anisakis,19 Termómetros,Otro"')
DESTINO_11 = ('"Desechado,Devuelto al proveedor,Reprocesado,'
              'Liberado tras evaluación,Pendiente de decisión,No aplica"')
VERIF_11 = '"Pendiente,Verificado OK,Requiere seguimiento"'


def _post_11(wb, fname, cambios):
    ws = wb['Acciones Correctivas']
    _lienzo(ws, F_CONT_11 + 10, NC_11 + 2)
    _titulo(ws, 'Registro de Acciones Correctivas', NC_11,
            'Año: ______    Establecimiento: ________________________    '
            'Responsable del sistema APPCC: ________________________')
    motor.cabecera(ws, 4, CAB_11, ANCHOS_11)

    for col in range(1, NC_11 + 1):
        cel = ws.cell(row=F0_11, column=col)
        if col == 1:
            motor.calculada(cel)
        elif col in (2, 12):
            motor.verde(cel, '@')
        elif col in (3, 4, 7, 13):
            motor.verde(cel, align='center')
        else:
            motor.verde(cel, align='left')
    motor.replicar_filas(ws, F0_11, F0_11, F1_11, ncols=NC_11, alto=30)
    for i in range(F0_11, F1_11 + 1):
        ws.cell(row=i, column=1).value = f'INC-{i - F0_11 + 1:03d}'

    motor.dv_lista(ws, TIPO_11, [motor.rango('C', F0_11, F1_11)],
                   'Tipo de incidencia', 'Elige un tipo de la lista.')
    motor.dv_lista(ws, ORIGEN_11, [motor.rango('D', F0_11, F1_11)],
                   'Registro de origen',
                   'Elige el registro donde se detectó la desviación. Es lo '
                   'que permite reconstruir el caso: sin él, la incidencia '
                   'queda suelta y no se puede verificar.')
    motor.dv_lista(ws, DESTINO_11, [motor.rango('G', F0_11, F1_11)],
                   'Destino del producto no conforme',
                   'Elige qué se hizo con el producto afectado. Sin la '
                   'disposición del producto no conforme, la acción correctiva '
                   'está incompleta (principio 5 del APPCC).')
    motor.dv_lista(ws, VERIF_11, [motor.rango('M', F0_11, F1_11)],
                   'Verificación',
                   'Elige el estado de verificación. El contador del final '
                   'compara por igualdad exacta, así que un texto libre no '
                   'suma.')
    motor.semaforo(ws, motor.rango('M', F0_11, F1_11),
                   extra_ok=('Verificado OK',),
                   extra_ambar=('Pendiente', 'Requiere seguimiento'))
    motor.semaforo(ws, motor.rango('G', F0_11, F1_11),
                   extra_ok=('Liberado tras evaluación',),
                   extra_ambar=('Reprocesado', 'Pendiente de decisión'),
                   extra_rojo=('Desechado', 'Devuelto al proveedor'))

    # Los ejemplos son los MISMOS casos que siembran los otros registros: el
    # cliente que siga el hilo desde una desviación hasta su cierre ve el
    # circuito completo. INC-001/002/003 vienen de 01, 02 y 10 (grupo A);
    # INC-004 de 07 (plagas), INC-005 de 17 (enfriamiento), INC-006 de 19
    # (termómetros) e INC-007 de 16 (regeneración).
    #
    # DOM-R2-03 / TEC-03 / COM-R2-05 (altas, ronda 2): este era el ÚNICO
    # registro sembrado del pack sin la marca «(ejemplo)» que exige la §1.5 —el
    # resto la lleva y el cliente aprende a fiarse de ella—, y son tres
    # incidencias completas, con lote, iniciales y «Verificado OK», en la hoja
    # que un inspector abre para ver el historial de no conformidades. Quien lo
    # imprimiera y lo archivase tenía en su carpeta APPCC tres desviaciones que
    # nunca ocurrieron en su local.
    #
    # DOM-R2-15 / TEC-03: además, 07, 16, 17 y 19 remitían a INC-004, INC-005 e
    # INC-006, que en esta hoja eran filas con el número y NADA más — y el
    # INC-004 lo usaban DOS ficheros para DOS sucesos distintos. Dos
    # desviaciones bajo el mismo número es justo lo que impide reconstruir un
    # caso, y era el ejemplo con el que el pack enseñaba a enganchar registros.
    # Aquí se desarrollan las cuatro y se renumera la del 16 a INC-007.
    motor.sembrar(ws, 5, {
        'B': '07/09/2026', 'C': 'Temperatura', 'D': '01 Temperaturas diario',
        'E': 'Cámara 1 a 6,5 °C en la lectura de la mañana del miércoles',
        'F': 'Producto refrigerado de la cámara 1 (lote L-2026-0915)',
        'G': 'Liberado tras evaluación',
        'H': 'Puerta mal cerrada tras el servicio de la noche',
        'I': 'Traslado del producto a la cámara 2, medición del centro del '
             'producto (4,1 °C) y ajuste del termostato',
        'J': 'Cartel de cierre de puerta y comprobación al cerrar añadida al '
             'checklist de fin de turno',
        'K': 'A.R.', 'L': '07/09/2026', 'M': 'Verificado OK'}, marca_col='E')
    motor.sembrar(ws, 6, {
        'B': '05/09/2026', 'C': 'Producto rechazado',
        'D': '02 Recepción temperaturas',
        'E': 'Guisantes congelados recibidos a -14 °C y con el envase roto',
        'F': 'Guisantes congelados, lote L-2026-0914',
        'G': 'Devuelto al proveedor',
        'H': 'Rotura de la cadena de frío en el transporte',
        'I': 'Rechazo de la partida completa y anotación en el albarán',
        'J': 'Aviso al proveedor y exigencia de registro de temperatura del '
             'vehículo en las próximas entregas',
        'K': 'A.R.', 'L': '05/09/2026', 'M': 'Verificado OK'}, marca_col='E')
    motor.sembrar(ws, 7, {
        'B': '15/09/2026', 'C': 'Otro', 'D': '10 Agua',
        'E': 'Agua turbia en el grifo de cocina; no se llegó a medir el cloro',
        'F': 'No aplica: se cerró el punto antes de usar el agua',
        'G': 'No aplica',
        'H': 'Obra en la red municipal de la calle',
        'I': 'Cierre del punto de consumo, uso de agua envasada y solicitud de '
             'análisis externo',
        'J': 'Suscripción a los avisos de corte del gestor de abastecimiento',
        'K': 'M.S.', 'M': 'Requiere seguimiento'}, marca_col='E')
    motor.sembrar(ws, 8, {
        'B': '18/09/2026', 'C': 'Plagas', 'D': '07 Plagas',
        'E': 'Un saco de harina roído en el almacén, con excrementos junto al '
             'palé',
        'F': 'Harina de trigo, lote L-2026-0930 (saco de 25 kg)',
        'G': 'Desechado',
        'H': 'Hueco bajo la puerta de servicio sin burlete',
        'I': 'Retirada y desecho del saco, limpieza y desinfección de la zona '
             'y aviso a la empresa DDD, que revisó las cinco estaciones',
        'J': 'Burlete en la puerta de servicio y palés separados 15 cm de la '
             'pared para poder inspeccionar',
        'K': 'M.S.', 'L': '19/09/2026', 'M': 'Verificado OK'}, marca_col='E')
    motor.sembrar(ws, 9, {
        'B': '07/09/2026', 'C': 'Temperatura', 'D': '17 Enfriamiento',
        'E': 'Estofado de ternera a 14 °C dos horas después de empezar a '
             'enfriar (límite: 10 °C)',
        'F': 'Estofado de ternera, 8 kg (elaboración del 07/09)',
        'G': 'Desechado',
        'H': 'Olla de 20 L metida entera en la cámara, sin porcionar',
        'I': 'Desecho de la elaboración completa y reparto del siguiente lote '
             'en bandejas de menos de 5 cm',
        'J': 'Bandejas planas obligatorias para todo enfriamiento y compra de '
             'dos bandejas gastronorm más',
        'K': 'M.S.', 'L': '08/09/2026', 'M': 'Verificado OK'}, marca_col='E')
    motor.sembrar(ws, 10, {
        'B': '01/09/2026', 'C': 'Otro', 'D': '19 Termómetros',
        'E': 'Termómetro de la cámara 2 con desviación de 1,8 °C en el hielo '
             'fundente (tolerancia ±1)',
        'F': 'Ninguno: se revisaron las lecturas del registro 01 desde la '
             'última verificación buena y estaban dentro de rango',
        'G': 'No aplica',
        'H': 'Golpe al mover la sonda; el equipo no admite recalibración',
        'I': 'Retirado de uso y sustituido el mismo día; relectura de la '
             'cámara 2 con la sonda S-01',
        'J': 'Verificación mensual fija el día 1 y funda de protección para '
             'las sondas de mano',
        'K': 'M.S.', 'L': '01/09/2026', 'M': 'Verificado OK'}, marca_col='E')
    motor.sembrar(ws, 11, {
        'B': '06/09/2026', 'C': 'Temperatura', 'D': '16 Cocción',
        'E': 'Estofado de ternera regenerado a 70 °C en el centro (límite: '
             '75 °C)',
        'F': 'Estofado de ternera, ración de servicio del 06/09',
        'G': 'Reprocesado',
        'H': 'Horno a media potencia y bandeja demasiado llena',
        'I': 'Vuelto a calentar hasta 79 °C, comprobado con sonda antes de '
             'servir',
        'J': 'Regeneración en tandas de media bandeja y sonda obligatoria '
             'antes de cada pase',
        'K': 'M.S.', 'L': '06/09/2026', 'M': 'Verificado OK'}, marca_col='E')

    f = (f'=COUNTIF(M{F0_11}:M{F1_11},"Pendiente")'
         f'+COUNTIF(M{F0_11}:M{F1_11},"Requiere seguimiento")')
    _contador(ws, F_CONT_11, NC_11, NC_11,
              'Incidencias abiertas (pendientes de verificar o en '
              'seguimiento):', f, rojo_si=f'=M{F_CONT_11}>0')

    _pie(ws, F_CONT_11 + 2, NC_11, extra=(
        'Toda incidencia tiene que tener acción correctiva (qué se hizo con lo '
        'que ya había pasado) Y acción preventiva (qué se cambia para que no '
        'vuelva a pasar). Si sólo hay la primera, la misma desviación reaparece '
        'el mes siguiente y eso es lo que un inspector lee en la serie.',
        'El «Destino del producto no conforme» es obligatorio siempre que haya '
        'producto afectado: desechado, devuelto, reprocesado o liberado tras '
        'evaluación. «Liberado» exige dejar escrito con qué criterio.',
        'El «Registro de origen» enlaza la incidencia con la hoja donde se '
        'detectó. Anota el nº de esta hoja (INC-0xx) también en esa otra: los '
        'registros 01 y 02 tienen columna para ello.',
        'Las siete primeras filas (INC-001 a INC-007) son EJEMPLOS: llevan '
        '«(ejemplo)» en la descripción y son los mismos casos que verás '
        'sembrados en los registros 01, 02, 07, 10, 16, 17 y 19. Bórralas '
        'antes de empezar a usar el registro: si las imprimes tal cual, estás '
        'archivando siete no conformidades que no ocurrieron en tu local.'))
    motor.IMPRESION[(fname, ws.title)] = (4, True, 'C5')  # TEC-08
    cambios.append('11: producto/lote afectado, destino del producto no '
                   'conforme, acción preventiva y registro de origen con DV; '
                   '40 filas numeradas INC-0xx y contador de abiertas (DOM-24); '
                   'los 7 ejemplos marcados «(ejemplo)» y desarrollados los '
                   'que citan 07, 16, 17 y 19 '
                   '(DOM-R2-03/DOM-R2-15/TEC-03/COM-R2-05)')

    motor.escribir_instrucciones(wb, 'Registro de Acciones Correctivas', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Registra CUALQUIER incidencia de seguridad alimentaria: '
              'desviaciones de temperatura, producto rechazado, fallos de '
              'limpieza, indicios de plagas, reclamaciones, alérgenos, '
              'caducidades.'),
        ('b', 'Cada fila viene numerada (INC-001, INC-002…). Anota ese mismo '
              'número en el registro donde detectaste la desviación: el 01 y el '
              '02 tienen columna «Nº incidencia» justo para eso.'),
        ('b', 'Las siete primeras filas son EJEMPLOS y lo dicen en la '
              'descripción: «(ejemplo)». Son los mismos casos que aparecen '
              'sembrados en los registros 01, 02, 07, 10, 16, 17 y 19, para '
              'que veas el circuito entero de una desviación. BÓRRALAS antes '
              'de usar la hoja: un libro de acciones correctivas con '
              'incidencias que no ocurrieron es un registro falsificado.'),
        ('b', 'El «Registro de origen» es un desplegable con los registros del '
              'pack. Sin él la incidencia queda suelta y no se puede '
              'reconstruir el caso.'),
        ('h', 'Las cuatro cosas que no pueden faltar'),
        ('b', 'Producto / lote afectado: qué quedó comprometido. Si no hubo '
              'producto afectado, escríbelo («No aplica»), pero no lo dejes en '
              'blanco.'),
        ('b', 'Destino del producto no conforme: desechado, devuelto al '
              'proveedor, reprocesado o liberado tras evaluación. Es el '
              'principio 5 del APPCC y es lo primero que se pregunta.'),
        ('b', 'Acción correctiva: qué se hizo con lo que ya había pasado.'),
        ('b', 'Acción preventiva: qué se cambia para que no vuelva a pasar. Es '
              'la que distingue un sistema vivo de una lista de excusas.'),
        ('h', 'Verificación'),
        ('b', 'Una incidencia no se cierra hasta que alguien distinto de quien '
              'la resolvió comprueba que la acción funcionó. Hasta entonces se '
              'queda en «Pendiente» o «Requiere seguimiento», y el contador del '
              'final de la hoja las suma en rojo.'),
        ('b', 'Un registro de acciones correctivas VACÍO no demuestra que todo '
              'va bien: demuestra que no se está vigilando. Tener incidencias '
              'anotadas y cerradas es lo que acredita que el sistema funciona.'),
        ('b', motor.CONSERVACION),
    ], cambios)


# ==========================================================================
# 12 — Análisis de peligros (el corazón del sistema)
# ==========================================================================
# Cuatro defectos, dos de ellos altos:
#   DOM-01/TEC-04/COM-04 — la columna «Registro» remitía a cuatro registros que
#     el pack NO entregaba, dos de ellos de PCC declarados (cocción y
#     enfriamiento). El plan se autoincumplía en su documento maestro.
#   DOM-06 — no existía el anisakis en todo el pack, siendo obligación con
#     sanción en cualquier local que sirva boquerones en vinagre, ceviche o
#     sushi (RD 1021/2022, art. 8.1, que derogó el RD 1420/2006, y Reg.
#     853/2004 Anexo III, Secc. VIII, Cap. III.D).
#   DOM-17/TEC-08/COM-20 — faltaban las columnas de medida preventiva y de
#     verificación (principio 6), Probabilidad y Gravedad no alimentaban nada, y
#     dos filas con la MISMA puntuación que otras se resolvían al revés. Dos
#     «PCC» tenían por límite crítico algo no medible.
#   DOM-26 — faltaban la descongelación (que el propio pack califica de falta
#     grave en la guía 15) y el huevo fresco sin tratamiento térmico.
CAB_12 = ['Fase del proceso', 'Tipo de peligro', 'Peligro identificado',
          'Medida preventiva', 'Probabilidad', 'Gravedad', 'Nivel de riesgo',
          '¿Es PCC?', 'Límite crítico', 'Registro (dónde se anota)',
          'Vigilancia (frecuencia y responsable)', 'Acción correctiva',
          'Verificación (qué, quién, cuándo)']
ANCHOS_12 = [24, 15, 40, 40, 13, 12, 15, 12, 44, 30, 38, 38, 40]
NC_12 = len(CAB_12)
F0_12 = 5
N_LIBRES_12 = 10

TIPO_12 = '"B (Biológico),Q (Químico),F (Físico)"'
NIVEL_12 = '"Alta,Media,Baja"'
PCC_12 = '"PCC,PPRo,NO"'

# 21 filas sobre SIETE fases (§3: «7 fases»). El orden es el del diagrama de
# flujo de un restaurante, que es como lo lee un inspector. La cifra sale de
# `len(PELIGROS_12)`: las Instrucciones y el `cambios` la leen de ahí, así que
# añadir una fila no deja ningún «18 peligros» rezagado dentro del fichero (en
# la landing sí hay que tocarlo a mano: pack-appcc.ts).
PELIGROS_12 = [
    ('RECEPCIÓN', 'B (Biológico)',
     'Multiplicación microbiana por temperatura elevada en producto '
     'refrigerado o congelado',
     'Proveedores homologados con transporte isotermo; medición con sonda en '
     'cada entrega',
     'Alta', 'Alta', 'PCC',
     'El límite legal de cada familia (hoja «Límites» del registro 02): '
     'pescado fresco 0-2 °C, carne picada ≤2 °C, aves ≤4 °C, canales ≤7 °C, '
     'congelados ≤-18 °C',
     '02 Recepción temperaturas',
     'En cada entrega; jefe de cocina o receptor designado',
     'Rechazar la partida y abrir la incidencia en el registro 11',
     'Revisión mensual del registro 02 por el responsable de APPCC'),
    ('RECEPCIÓN', 'Q (Químico)',
     'Producto sin etiquetado de alérgenos o con etiqueta ilegible',
     'Fichas técnicas de proveedor y verificación documental de cada albarán',
     'Media', 'Alta', 'PPRo',
     '100 % de las referencias recibidas con etiqueta legible que declare los '
     '14 alérgenos',
     '05 Recepción mercancías',
     'En cada entrega; receptor',
     'Rechazar la referencia y reclamar la ficha técnica al proveedor',
     'Revisión trimestral de las fichas técnicas; responsable de APPCC'),
    ('RECEPCIÓN', 'F (Físico)',
     'Envase dañado o presencia de cuerpos extraños',
     'Inspección visual del embalaje antes de descargar',
     'Baja', 'Media', 'NO',
     'Envase íntegro, limpio y sin hinchazón',
     '05 Recepción mercancías',
     'En cada entrega; receptor',
     'Rechazar la unidad afectada',
     'Revisión mensual del registro 05'),
    ('ALMACENAMIENTO', 'B (Biológico)',
     'Rotura de la cadena de frío en cámaras, congeladores y exposición',
     'Mantenimiento preventivo de los equipos y sondas verificadas',
     'Media', 'Alta', 'PCC',
     'Refrigeración 0-4 °C · congelación ≤-18 °C · exposición caliente ≥65 °C',
     '01 Temperaturas diario',
     'Dos lecturas diarias por equipo (mañana y tarde); cocinero de turno',
     'Trasladar el producto, evaluar su temperatura en el centro y desechar si '
     'procede; abrir el registro 11',
     'Verificación mensual de las sondas (registro 19) y revisión del 01'),
    ('ALMACENAMIENTO', 'B (Biológico)',
     'Contaminación cruzada entre crudo y cocinado en cámara',
     'Separación física, crudos en la balda inferior, producto tapado y '
     'etiquetado',
     'Media', 'Alta', 'PPRo',
     '100 % del producto tapado y etiquetado, con los crudos por debajo del '
     'producto listo para consumo',
     '04 Limpieza diaria (bloque semanal)',
     'Revisión semanal de cámaras; jefe de cocina',
     'Reorganizar la cámara y desechar lo que haya podido contaminarse',
     'Auditoría interna trimestral; responsable de APPCC'),
    ('ALMACENAMIENTO', 'Q (Químico)',
     'Producto caducado o sin identificar en stock',
     'FIFO y etiquetado con fecha de elaboración y de consumo preferente',
     'Media', 'Media', 'NO',
     'Cero productos caducados o sin etiqueta en cámara y almacén',
     '06 Trazabilidad',
     'Revisión semanal; almacenero',
     'Retirar el producto y registrar la retirada',
     'Revisión mensual del registro 06'),
    # DOM-R2-09 (ronda 2): el pack VENDE los registros 07 (plagas), 10 (agua) y
    # 09 (aceite) y el análisis no contemplaba ninguno de los tres peligros. El
    # 11 los ofrecía los tres en el desplegable «Registro de origen», así que el
    # cliente podía abrir una incidencia procedente de un registro cuyo peligro
    # no estaba analizado: para un inspector, un análisis incompleto justo en el
    # documento que el propio fichero llama «el corazón del sistema».
    ('ALMACENAMIENTO', 'B (Biológico)',
     'Contaminación por plagas y vectores (roedores, insectos, aves) en '
     'almacén, cocina y cámara de residuos',
     'Programa DDD con empresa inscrita en el ROESB, barreras físicas '
     '(mosquiteras, burletes), residuos cerrados y revisión de estaciones',
     'Media', 'Alta', 'PPRo',
     'Cero indicios en la revisión (excrementos, roeduras, insectos vivos) y '
     'estaciones de cebo íntegras y sin consumo',
     '07 Control de plagas (DDD)',
     'Revisión visual mensual y actuación de la empresa DDD según contrato; '
     'responsable de APPCC',
     'Avisar a la empresa DDD, desechar el producto afectado y sellar la vía '
     'de entrada; abrir incidencia en el registro 11',
     'Revisión del informe de la empresa DDD y del registro 07; anual'),
    ('DESCONGELACIÓN', 'B (Biológico)',
     'Multiplicación microbiana por descongelar a temperatura ambiente o bajo '
     'el grifo',
     'Planificación de la producción y descongelación programada en cámara',
     'Alta', 'Alta', 'PCC',
     'En refrigeración a ≤4 °C, tapado y sobre bandeja con rejilla; consumo en '
     # DOM-R2-11: el plan decía «≤24 h desde el FIN» y su registro (17) mide
     # 24 h desde el INICIO. Dos papeles del mismo pack con dos límites
     # críticos distintos para el mismo PCC es lo que un inspector usa para
     # tumbar el sistema. Se elige el criterio del 17 —el que acota el tiempo
     # TOTAL en la zona de riesgo y el que las columnas ya pueden medir.
     '≤24 h desde el INICIO de la descongelación (es lo que mide el registro '
     '17: acota el tiempo total en la zona de riesgo)',
     '17 Enfriamiento y descongelación',
     'En cada descongelación; cocinero de turno',
     'Desechar el producto descongelado fuera de cámara; abrir el registro 11',
     'Revisión semanal del registro 17; jefe de cocina'),
    ('PREPARACIÓN', 'B (Biológico)',
     'Anisakis en pescado destinado a consumo en crudo, marinado, escabechado '
     'o salazón ligera (boquerones en vinagre, ceviche, sushi, tataki)',
     'Compra de pescado eviscerado y congelación preventiva ANTES de la '
     'elaboración',
     'Alta', 'Alta', 'PCC',
     '-20 °C durante al menos 24 h en todas las partes de la pieza, o -35 °C '
     'durante al menos 15 h (RD 1021/2022, art. 8.1, que derogó el '
     'RD 1420/2006, y Rgto. (CE) 853/2004, Anexo III, Secc. VIII, Cap. III.D)',
     '18 Congelación preventiva de anisakis',
     'En cada lote destinado a crudo o semicrudo; jefe de cocina',
     'No servir la pieza: repetir la congelación completa o desecharla',
     'Revisión mensual del registro 18 y de la temperatura del congelador '
     '(registro 01)'),
    ('PREPARACIÓN', 'B (Biológico)',
     'Salmonella por huevo fresco en elaboraciones sin tratamiento térmico '
     '(mayonesas, salsas, postres)',
     'Uso de ovoproducto pasteurizado en toda elaboración que no alcance los '
     '75 °C',
     'Alta', 'Alta', 'PCC',
     # DOM-R2-10: el registro asignado era el 16 (Cocción y regeneración), que
     # sólo tiene columnas de temperatura en el centro, tiempo y sonda: una
     # mayonesa hecha con ovoproducto pasteurizado no deja ni una línea ahí, y
     # ante «enséñeme los registros de vigilancia de este PCC» no había nada
     # que enseñar. Lo que SÍ se puede acreditar es la compra: verificación
     # documental del albarán y la etiqueta en el registro 05. Las
     # elaboraciones que sí alcanzan los 75 °C siguen en el 16.
     'Ovoproducto pasteurizado en toda elaboración que no alcance ≥75 °C en el '
     'centro; si los alcanza, se acredita en el registro 16',
     '05 Recepción mercancías (compra de ovoproducto: albarán y etiquetado)',
     'En cada entrega de ovoproducto y en cada elaboración en frío; receptor y '
     'cocinero de turno',
     'Desechar la elaboración y rehacerla con ovoproducto pasteurizado',
     'Revisión mensual de las compras de ovoproducto (registro 05) y del 16'),
    ('PREPARACIÓN', 'B (Biológico)',
     'Agua de consumo no potable en lavado de vegetales, cocción, hielo o '
     'limpieza de superficies',
     'Suministro de red municipal; con pozo o depósito propio, cloración y '
     'control del cloro residual libre y del aspecto',
     'Media', 'Alta', 'PPRo',
     'Cloro residual libre entre 0,2 y 1,0 mg/L y agua sin turbidez, color ni '
     'olor extraños (RD 3/2023)',
     '10 Control de agua potable',
     'Diaria con depósito o pozo propio, semanal con red municipal; '
     'responsable de APPCC',
     'Cerrar el punto de consumo, usar agua envasada y solicitar análisis '
     'externo; abrir incidencia en el registro 11',
     'Análisis externo según la periodicidad del RD 3/2023 y revisión mensual '
     'del registro 10'),
    ('PREPARACIÓN', 'B (Biológico)',
     'Contaminación cruzada por utensilios, tablas y superficies',
     'Tablas y cuchillos por color, limpieza y desinfección entre usos y '
     'formación del personal',
     'Media', 'Alta', 'PPRo',
     'Utensilio limpio y desinfectado entre crudo y cocinado; 100 % de las '
     'tareas del plan L+D marcadas',
     '04 Limpieza diaria',
     'Continua durante el servicio; jefe de cocina',
     'Retirar y lavar el utensilio; rehacer la elaboración si hubo contacto',
     'Revisión semanal del registro 04 y test de superficie trimestral'),
    ('PREPARACIÓN', 'Q (Químico)',
     'Alérgeno transferido por utensilio, freidora o agua de cocción '
     'compartida',
     'Utensilios, tabla y zona exclusivos; freidora, aceite y agua sin '
     'compartir (protocolo de la ficha 14)',
     # DOM-R2-10: el registro asignado era el 08, que es la matriz de la CARTA
     # —una fila por plato, no por comanda— y no tiene fecha, ni turno, ni
     # firma: la vigilancia «en cada comanda con alergia declarada» no se podía
     # documentar en ninguna parte. Se degrada a PPRo, que es lo coherente con
     # la regla de decisión que enuncia el propio fichero: el control es de
     # prerrequisito (utensilios exclusivos, formación) y no hay límite medible
     # en tiempo real. Si una comanda se desvía, se abre incidencia en el 11.
     'Media', 'Alta', 'PPRo',
     'Matriz de alérgenos revisada y utensilios, aceite y agua exclusivos en '
     'el 100 % de las comandas con alergia declarada; el personal que las '
     'elabora, formado',
     '08 Matriz de alérgenos + 11 Acciones correctivas (si hay desviación)',
     'En cada comanda con alergia declarada; jefe de partida',
     'Desechar el plato y elaborarlo de nuevo desde cero; abrir incidencia en '
     'el registro 11',
     'Revisión mensual de la matriz y formación anual en alérgenos'),
    ('PREPARACIÓN', 'F (Físico)',
     'Restos metálicos de estropajo, plásticos o cuerpos extraños',
     'Estropajos de uso alimentario y revisión del estado de los utensilios',
     'Baja', 'Alta', 'NO',
     'Utensilios íntegros, sin filamentos ni piezas sueltas',
     '04 Limpieza diaria',
     'Diaria, al montar el servicio; cocinero de turno',
     'Sustituir el utensilio y desechar el producto afectado',
     'Revisión mensual del estado de utensilios; jefe de cocina'),
    ('COCCIÓN', 'B (Biológico)',
     'Supervivencia de patógenos por temperatura insuficiente en el centro del '
     'producto',
     'Sonda verificada y tiempos de cocción por ficha técnica de cada '
     'elaboración',
     'Media', 'Alta', 'PCC',
     '≥75 °C en el centro del producto; en regeneración, alcanzar los 75 °C en '
     'menos de una hora',
     '16 Cocción y regeneración',
     'En cada tanda; cocinero de turno',
     'Prolongar la cocción hasta alcanzar el límite; si no se alcanza, '
     'desechar',
     'Verificación mensual de la sonda (registro 19) y revisión del 16'),
    ('COCCIÓN', 'Q (Químico)',
     'Compuestos polares por degradación del aceite de fritura',
     'Filtrado diario, reposición de nivel, control de la temperatura de la '
     'freidora y test periódico de compuestos polares',
     'Media', 'Alta', 'PCC',
     'Máximo 25 % de compuestos polares (Orden de 26 de enero de 1989) y '
     'temperatura de fritura no superior a 180 °C',
     '09 Control de aceite de fritura',
     'Un test por freidora y semana en uso normal, diario si se fríe empanado '
     'o pescado a diario; jefe de cocina',
     'Cambiar el aceite y entregarlo a gestor autorizado; abrir incidencia en '
     'el registro 11',
     'Revisión mensual del registro 09 y del justificante de retirada'),
    ('MANTENIMIENTO Y ENFRIAMIENTO', 'B (Biológico)',
     'Multiplicación microbiana en mantenimiento caliente por debajo de 65 °C',
     'Baño maría y vitrinas calientes con termostato revisado; reposición en '
     'tandas pequeñas',
     'Media', 'Alta', 'PCC',
     '≥65 °C en todo el producto expuesto, sin límite por arriba',
     '01 Temperaturas diario (exposición caliente)',
     'Al montar el servicio y cada 2 h; cocinero de turno',
     'Regenerar a ≥75 °C, o desechar si lleva más de 2 h por debajo de 65 °C',
     'Revisión semanal del registro 01; jefe de cocina'),
    ('MANTENIMIENTO Y ENFRIAMIENTO', 'B (Biológico)',
     'Enfriamiento lento a través de la zona de peligro (60-10 °C)',
     'Abatidor de temperatura o porcionado en recipientes de poca altura, '
     'destapados hasta enfriar',
     'Media', 'Alta', 'PCC',
     'De 60 °C a 10 °C en menos de 2 horas',
     '17 Enfriamiento y descongelación',
     'En cada elaboración que se enfríe para conservar; cocinero de turno',
     'Desechar la elaboración si supera las 2 horas',
     'Revisión semanal del registro 17; jefe de cocina'),
    ('MANTENIMIENTO Y ENFRIAMIENTO', 'B (Biológico)',
     'Vida útil superada en elaborados propios',
     'Etiquetado de todo elaborado con fecha de elaboración y de consumo',
     'Media', 'Media', 'NO',
     'Todo elaborado etiquetado y consumido dentro del plazo de su ficha',
     '06 Trazabilidad',
     'Diaria, al montar y al cerrar; jefe de cocina',
     'Retirar el producto y registrar la retirada',
     'Revisión semanal del registro 06'),
    ('SERVICIO', 'B (Biológico)',
     'Exposición prolongada a temperatura ambiente en línea de servicio o '
     'buffet',
     'Reposición en tandas pequeñas y control horario de lo expuesto',
     'Media', 'Media', 'NO',
     'Máximo 2 horas a temperatura ambiente',
     '01 Temperaturas diario (exposición caliente)',
     'Control horario en cada servicio; responsable de sala',
     'Retirar el alimento expuesto',
     'Revisión semanal del registro 01'),
    ('SERVICIO', 'Q (Químico)',
     'Alérgeno no declarado al cliente en la carta o de palabra',
     'Carta de alérgenos actualizada tras cada cambio de receta y personal de '
     'sala formado',
     'Media', 'Alta', 'PPRo',
     'Carta de alérgenos revisada y firmada en los últimos 30 días o tras cada '
     'cambio de receta',
     '08 Matriz de alérgenos',
     'Revisión mensual y tras cada cambio de carta; responsable de APPCC',
     'Actualizar la matriz, comunicarlo a sala y revisar las comandas del día',
     'Auditoría interna trimestral y formación anual en alérgenos'),
]

F1_12 = F0_12 + len(PELIGROS_12) - 1            # última fila con contenido
FL1_12 = F1_12 + N_LIBRES_12                    # última fila libre
F_CONT_12 = FL1_12 + 2

# TEC-08: Probabilidad y Gravedad no alimentaban NADA y las decisiones de PCC
# se contradecían entre sí (seis filas Media/Alta → PCC y dos idénticas → NO).
# Aquí el nivel de riesgo se CALCULA: Alta=3, Media=2, Baja=1, y el producto
# decide. Se escribe tres veces el producto porque Excel no tiene variables
# locales; pycel lo evalúa igual.
def _f_riesgo(fila):
    p = f'IF($E{fila}="Alta",3,IF($E{fila}="Media",2,1))'
    g = f'IF($F{fila}="Alta",3,IF($F{fila}="Media",2,1))'
    return (f'=IF(OR($E{fila}="",$F{fila}=""),"",'
            f'IF({p}*{g}>=9,"Crítico",'
            f'IF({p}*{g}>=6,"Alto",'
            f'IF({p}*{g}>=3,"Medio","Bajo"))))')


def _post_12(wb, fname, cambios):
    ws = wb['Análisis Peligros']
    _lienzo(ws, F_CONT_12 + 14, NC_12 + 2)
    _titulo(ws, 'Análisis de Peligros — Plan APPCC', NC_12,
            'Establecimiento: ________________________    '
            'Fecha: ___/___/______    Revisión nº: _____    '
            'Responsable del sistema: ________________________')
    motor.cabecera(ws, 4, CAB_12, ANCHOS_12)

    for col in range(1, NC_12 + 1):
        cel = ws.cell(row=F0_12, column=col)
        if col == 7:
            motor.calculada(cel)
        elif col in (5, 6, 8):
            motor.verde(cel, align='center')
        else:
            motor.verde(cel, align='left')
    f_r = _f_riesgo(F0_12)
    ws.cell(row=F0_12, column=7).value = f_r
    motor.reg(ws, f'G{F0_12}', f_r)
    motor.replicar_filas(ws, F0_12, F0_12, FL1_12, ncols=NC_12)

    for i, datos in enumerate(PELIGROS_12):
        fila = F0_12 + i
        fase, tipo, peligro, prev, prob, grav, pcc, lim, reg, vig, ac, ver = datos
        for col, txt in ((1, fase), (2, tipo), (3, peligro), (4, prev),
                         (5, prob), (6, grav), (8, pcc), (9, lim), (10, reg),
                         (11, vig), (12, ac), (13, ver)):
            ws.cell(row=fila, column=col).value = txt
        ws.row_dimensions[fila].height = _alto(
            max((peligro, prev, lim, vig, ac, ver), key=len), 42, minimo=52)
    for fila in range(F1_12 + 1, FL1_12 + 1):
        ws.row_dimensions[fila].height = 30

    # TEC-25: los tres desplegables de v1.1 cubrían SÓLO las diez filas en
    # blanco y dejaban sin validación las trece de contenido, que son
    # justamente las que el cliente edita para adaptarlas a su local.
    def todas(col):
        return motor.rango(col, F0_12, FL1_12)

    motor.dv_lista(ws, TIPO_12, [todas('B')], 'Tipo de peligro',
                   'Elige B (biológico), Q (químico) o F (físico).')
    motor.dv_lista(ws, NIVEL_12, [todas('E'), todas('F')],
                   'Probabilidad y gravedad',
                   'Elige Alta, Media o Baja. El «Nivel de riesgo» se calcula '
                   'solo a partir de estas dos columnas: si escribes texto '
                   'libre, la fórmula lo tratará como «Baja».')
    motor.dv_lista(ws, PCC_12, [todas('H')], '¿Es PCC?',
                   'PCC = punto de control crítico · PPRo = prerrequisito '
                   'operativo · NO = se cubre con los prerrequisitos '
                   'generales.')
    motor.semaforo(ws, todas('G'),
                   extra_ok=('Bajo',), extra_ambar=('Medio',),
                   extra_rojo=('Crítico', 'Alto'))

    _contador(ws, F_CONT_12, NC_12, 8,
              'Peligros identificados en el análisis:',
              f'=COUNTIF(H{F0_12}:H{FL1_12},"PCC")'
              f'+COUNTIF(H{F0_12}:H{FL1_12},"PPRo")'
              f'+COUNTIF(H{F0_12}:H{FL1_12},"NO")')
    _contador(ws, F_CONT_12 + 1, NC_12, 8,
              'Puntos de control crítico (PCC):',
              f'=COUNTIF(H{F0_12}:H{FL1_12},"PCC")')
    _contador(ws, F_CONT_12 + 2, NC_12, 8,
              'Prerrequisitos operativos (PPRo):',
              f'=COUNTIF(H{F0_12}:H{FL1_12},"PPRo")')

    _pie(ws, F_CONT_12 + 4, NC_12, extra=(
        'PCC = Punto de Control Crítico: fase en la que se puede aplicar un '
        'control esencial para eliminar un peligro o reducirlo a un nivel '
        'aceptable, con un límite crítico MEDIBLE y vigilado en tiempo real. '
        'PPRo = Programa de Prerrequisitos Operativo: el peligro es relevante '
        'pero se controla con proveedores, limpieza, formación o etiquetado, y '
        'no hay un límite medible en el momento.',
        # DOM-R2-18: la frase afirmaba lo contrario de lo que hace la tabla.
        # Nueve filas puntúan «Alto» y cinco son PCC mientras cuatro son PPRo;
        # lo que las separa no es la puntuación, es si hay una fase posterior
        # que elimine el peligro y si el límite es medible en tiempo real.
        'Cómo se decide, y por qué dos filas con la misma puntuación pueden '
        'acabar en PCC o en PPRo: nivel Alto o Crítico y sin fase posterior que elimine '
        'el peligro → PCC; nivel Alto o Crítico pero controlado por un '
        'prerrequisito → PPRo; nivel Bajo o Medio → NO.',
        'Cada fila de esta tabla tiene detrás un registro que se entrega en '
        'este mismo pack: recepción → 02, almacenamiento y exposición caliente '
        '→ 01, plagas → 07, agua → 10, descongelación y enfriamiento → 17, '
        'anisakis → 18, ovoproducto → 05, cocción y regeneración → 16, aceite '
        'de fritura → 09, limpieza → 04, trazabilidad → 06, alérgenos → 08, y '
        'toda desviación → 11. Un PCC sin registro no acredita nada.',
        'Revisa este análisis al menos una vez al año, y siempre que cambies '
        'la carta, un proceso, un equipo o un proveedor. Anota el número de '
        'revisión en la cabecera.'))
    motor.IMPRESION[(fname, ws.title)] = (4, True, 'C5')  # TEC-08

    fases = []
    for datos in PELIGROS_12:
        if datos[0] not in fases:
            fases.append(datos[0])
    cambios.append(
        f'12: {len(PELIGROS_12)} peligros en {len(fases)} fases con medida '
        'preventiva, vigilancia y verificación; nivel de riesgo calculado; '
        'descongelación, huevo fresco y anisakis añadidos; columna «Registro» '
        'apuntando sólo a ficheros del pack; DV sobre todas las filas '
        '(DOM-01/DOM-06/DOM-17/DOM-26/TEC-04/TEC-08/TEC-25/COM-04/COM-20)')
    _instrucciones_12(wb, cambios, len(PELIGROS_12), fases)


def _instrucciones_12(wb, cambios, n, fases):
    motor.escribir_instrucciones(wb, 'Análisis de Peligros — Plan APPCC', [
        ('h', 'Qué es este documento'),
        ('p', 'Es el corazón del sistema APPCC: identifica los peligros de cada '
              'fase del proceso, decide cuáles son puntos de control crítico y '
              'dice, para cada uno, qué límite no se puede rebasar, quién lo '
              'vigila, qué se hace si se rebasa, dónde se anota y quién '
              'comprueba que todo eso se cumple.'),
        ('b', f'Trae {n} peligros analizados en {len(fases)} fases: '
              + ', '.join(f.capitalize() for f in fases) + '.'),
        ('b', 'Adáptalo a TU local: todas las celdas verdes son editables y '
              'los desplegables funcionan en todas las filas, incluidas las que '
              'vienen rellenadas.'),
        ('h', 'Cómo se calcula el nivel de riesgo'),
        ('b', 'Probabilidad y Gravedad son desplegables (Alta / Media / Baja). '
              'El «Nivel de riesgo» se calcula solo: Alta=3, Media=2, Baja=1, '
              'y el producto de las dos da Crítico (9), Alto (6), Medio (3-4) '
              'o Bajo (1-2). Se pinta con semáforo.'),
        ('b', 'La decisión de PCC sigue una regla, no el criterio del día: '
              'riesgo Alto o Crítico sin fase posterior que elimine el peligro '
              '→ PCC; riesgo Alto o Crítico controlado por un prerrequisito '
              '(proveedores, limpieza, formación) y sin límite medible en '
              'tiempo real → PPRo; riesgo Bajo o Medio → NO.'),
        ('b', 'Por eso el etiquetado de alérgenos en recepción y la carta de '
              'alérgenos en servicio figuran como PPRo y no como PCC: «etiqueta '
              'completa» y «carta actualizada» no son límites críticos '
              'medibles. El control real del alérgeno está en cocina, y ahí sí '
              'es PCC.'),
        ('h', 'Los 7 principios del APPCC y dónde vive cada uno'),
        ('b', '1. Análisis de peligros → columnas «Peligro identificado», '
              '«Probabilidad», «Gravedad» y «Nivel de riesgo».'),
        ('b', '2. Determinación de los PCC → columna «¿Es PCC?».'),
        ('b', '3. Límites críticos → columna «Límite crítico».'),
        ('b', '4. Vigilancia → columna «Vigilancia (frecuencia y '
              'responsable)» y los registros del pack.'),
        ('b', '5. Medidas correctivas → columna «Acción correctiva» y el '
              'registro 11.'),
        ('b', '6. Verificación → columna «Verificación (qué, quién, cuándo)».'),
        ('b', '7. Documentación y registro → columna «Registro (dónde se '
              'anota)», que apunta sólo a ficheros que este pack entrega.'),
        ('h', 'Marco normativo'),
        ('b', 'Reg. (CE) 852/2004, art. 5 — obligación de implantar '
              'procedimientos basados en los principios del APPCC.'),
        ('b', 'RD 1021/2022, de 13 de diciembre — normas sanitarias de '
              'elaboración y comercialización de comidas preparadas. Derogó al '
              'RD 3484/2000 y ya no fija temperaturas ni tiempos: los establece '
              'y justifica el operador en su propio APPCC, que es lo que hace '
              'esta tabla.'),
        ('b', 'Reg. (CE) 178/2002, arts. 18 y 19 — trazabilidad, retirada e '
              'información inmediata a la autoridad competente.'),
        ('b', 'Reg. (UE) 1169/2011 y RD 126/2015 — información alimentaria '
              'facilitada al consumidor, incluidos los alérgenos.'),
        ('b', 'RD 1021/2022, art. 8.1 —que derogó el RD 1420/2006— y '
              'Rgto. (CE) 853/2004, Anexo III, Secc. VIII, Cap. III.D — '
              'congelación preventiva del pescado destinado a consumo en crudo '
              '(anisakis). El art. 8.2 añade informar al consumidor con cartel '
              'o carta-menú.'),
        ('b', 'Reg. (CE) 2073/2005 — criterios microbiológicos aplicables a '
              'los productos alimenticios.'),
        ('b', 'Ley 17/2011, de seguridad alimentaria y nutrición — régimen '
              'sancionador (arts. 50-52).'),
        ('b', motor.CONSERVACION),
    ], cambios)


# ==========================================================================
# 13 — Checklist de higiene personal (cartel de vestuario)
# ==========================================================================
# DOM-14: el pack exigía «Carné/certificado de manipulador de alimentos
# vigente». El carné se suprimió en España con el RD 109/2010: ya no existe ni
# carné oficial ni caducidad, y perpetuar el mito lleva al cliente a pagar
# «carnés» y a creerse en regla por tenerlos. Lo vigente es que la EMPRESA
# acredite la formación de sus manipuladores (Reg. (CE) 852/2004, Anexo II,
# Cap. XII).
# TEC-29: la línea de firma medía 76 caracteres sobre un ancho de 67 y no había
# área de impresión, así que el cartel salía cortado por el remate. Aquí la
# línea se parte en dos y el área la fija el motor sobre el contenido real.
NC_13 = 3
HIGIENE_13 = [
    ('sec', 'INDUMENTARIA'),
    ('i', 'Uniforme limpio y de uso exclusivo para el trabajo'),
    ('i', 'Gorro o cofia que cubra todo el cabello'),
    ('i', 'Calzado cerrado, antideslizante y limpio'),
    ('i', 'Sin joyas, relojes, pulseras ni piercings visibles'),
    ('i', 'Delantal limpio; cambiarlo en cuanto se ensucie'),
    ('sec', 'MANOS'),
    ('i', 'Uñas cortas, limpias, sin esmalte ni uñas postizas'),
    ('i', 'Lavado de manos al entrar en la cocina'),
    ('i', 'Lavado de manos después de ir al baño'),
    ('i', 'Lavado de manos al cambiar de tarea o de alimento'),
    ('i', 'Lavado de manos después de tocar basura, cajas o dinero'),
    ('i', 'Guantes si hay heridas; cambiarlos al cambiar de tarea'),
    ('i', 'Heridas cubiertas con apósito azul impermeable'),
    ('sec', 'CONDUCTA Y SALUD'),
    ('i', 'No comer, beber, mascar chicle ni fumar en la zona de trabajo'),
    ('i', 'No toser ni estornudar sobre los alimentos'),
    ('i', 'No tocarse cara, pelo ni nariz mientras se manipula comida'),
    ('i', 'Avisar al encargado ante vómitos, diarrea, fiebre o infección de '
          'la piel'),
    ('i', 'No trabajar con síntomas gastrointestinales hasta 48 h después de '
          'que cesen'),
    ('i', 'Ropa de calle y ropa de trabajo separadas en la taquilla'),
    ('sec', 'FORMACIÓN Y REGISTRO'),
    ('i', 'Formación en higiene alimentaria acreditada por la empresa '
          '(Reg. (CE) 852/2004, Anexo II, Cap. XII; el carné oficial se '
          'suprimió con el RD 109/2010)'),
    ('i', 'Formación específica en alérgenos y protocolo de actuación '
          '(ficha 14)'),
    ('i', 'Formación básica en APPCC recibida y registrada'),
    ('i', 'Alta en el registro de formación del establecimiento (BONUS-01)'),
]


def _post_13(wb, fname, cambios):
    ws = wb['Higiene Personal']
    _lienzo(ws, 60, 6)
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 12

    cel = ws.cell(row=2, column=2,
                  value='NORMAS DE HIGIENE PERSONAL — Zona de manipulación')
    cel.font = Font(bold=True, size=15)
    cel.alignment = Alignment(vertical='center')
    ws.merge_cells('B2:C2')
    ws.row_dimensions[2].height = 24
    cel = ws.cell(row=3, column=2,
                  value='Reg. (CE) 852/2004, Anexo II, Cap. VIII. '
                        'Comprueba estos puntos ANTES de entrar a manipular.')
    cel.font = Font(size=9, italic=True)

    fila, n_items = 5, 0
    for tipo, texto in HIGIENE_13:
        if tipo == 'sec':
            fila += 1 if n_items else 0
            c = ws.cell(row=fila, column=2, value=texto)
            c.font = Font(bold=True, size=11)
            c.fill = PatternFill('solid', fgColor=motor.SECCION)
            c.alignment = Alignment(vertical='center')
            cc = ws.cell(row=fila, column=3)
            cc.fill = PatternFill('solid', fgColor=motor.SECCION)
            ws.row_dimensions[fila].height = 20
            fila += 1
            continue
        c = ws.cell(row=fila, column=2, value=texto)
        c.font = Font(size=11)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = motor.BORDE
        box = ws.cell(row=fila, column=3, value='☐')
        box.font = Font(size=14)
        box.alignment = Alignment(horizontal='center', vertical='center')
        box.border = motor.BORDE
        ws.row_dimensions[fila].height = _alto(texto, 60, minimo=20)
        n_items += 1
        fila += 1

    fila += 1
    for txt in ('Responsable de la verificación: ____________________________',
                'Fecha: ___/___/______      Firma: __________________________'):
        c = ws.cell(row=fila, column=2, value=txt)
        c.font = Font(size=11, bold=True)
        fila += 1

    fila += 1
    motor.nota(ws, fila,
               'Imprime este cartel y cuélgalo en el vestuario y en la puerta '
               'de acceso a la cocina. El encargado lo verifica semanalmente y '
               'lo firma.', ncols=NC_13, ancho_car=78)
    motor.nota(ws, fila + 1,
               'El carné oficial de manipulador de alimentos NO existe desde '
               'el RD 109/2010: lo que la inspección comprueba es que la '
               'EMPRESA acredite la formación en higiene de cada trabajador '
               '(Reg. (CE) 852/2004, Anexo II, Cap. XII). Registra esa '
               'formación en el BONUS-01.', ncols=NC_13, ancho_car=78)
    motor.nota(ws, fila + 2, motor.CONSERVACION, ncols=NC_13, ancho_car=78)
    motor.nota(ws, fila + 3, motor.MARCA, ncols=NC_13, ancho_car=78)
    motor.IMPRESION[(fname, ws.title)] = (None, False, False)
    cambios.append(f'13: {n_items} puntos en 4 bloques, formación acreditada '
                   'en vez del carné suprimido y remate de firma partido para '
                   'que el área de impresión no lo corte (DOM-14/TEC-29)')

    motor.escribir_instrucciones(wb, 'Checklist de Higiene Personal', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Imprímela en A4 vertical y cuélgala en el vestuario y en la '
              'puerta de acceso a la cocina.'),
        ('b', 'Todo el personal debe cumplir estos puntos ANTES de entrar en '
              'la zona de manipulación.'),
        ('b', 'El encargado verifica semanalmente, marca las casillas y firma '
              'al pie. Archiva la hoja firmada.'),
        ('h', 'El carné oficial de manipulador ya no existe'),
        ('p', 'El carné oficial de manipulador de alimentos se suprimió en '
              'España con el RD 109/2010, que derogó el RD 202/2000. No hay '
              'carné, no hay caducidad y no hay que renovar nada en ninguna '
              'ventanilla. Lo que exige la norma vigente (Reg. (CE) 852/2004, '
              'Anexo II, Cap. XII) es que la EMPRESA garantice y pueda '
              'acreditar que sus manipuladores están formados y supervisados. '
              'Esa formación la registras en el BONUS-01.'),
        ('h', 'Bajas por enfermedad'),
        ('b', 'Quien tenga vómitos, diarrea, fiebre, lesiones cutáneas '
              'infectadas o supuración no puede manipular alimentos. La vuelta '
              'al puesto se hace 48 horas después de que cesen los síntomas.'),
        ('b', 'El trabajador tiene la obligación de comunicarlo y la empresa la '
              'de recolocarlo o darle de baja: es responsabilidad compartida.'),
        ('b', motor.CONSERVACION),
    ], cambios)


# ==========================================================================
# 14 — Fichas de los 14 alérgenos (cartel de cocina y sala)
# ==========================================================================
# DOM-27/TEC-09: el «protocolo» que se cuelga en cocina cabía en una línea, en
# una celda combinada SIN ajuste de texto —así que además salía cortada al
# imprimir— y omitía justo lo que evita un incidente: guantes y utensilios
# exclusivos, freidora y agua sin compartir, la frase honesta cuando no se
# puede garantizar, y qué hacer si alguien empieza una reacción.
NC_14 = 4
ALERGENOS_14 = [
    ('Gluten', 'Cereales con gluten: trigo, centeno, cebada, avena, espelta, '
     'kamut y sus derivados',
     'Pan, pasta, bollería, rebozados, cerveza, salsas con harina, cuscús, '
     'seitán'),
    ('Crustáceos', 'Cangrejos, langostinos, gambas, bogavante, nécora, cigalas '
     'y derivados',
     'Paella, salpicón de marisco, croquetas de marisco, sopas de pescado, '
     'surimi'),
    ('Huevos', 'Huevos y derivados de cualquier ave',
     'Tortillas, rebozados, mayonesa, pasta fresca, merengue, bizcochos, '
     'helados, flanes'),
    ('Pescado', 'Todas las especies de pescado y sus derivados',
     'Paella, caldos, sopas, surimi, salsa Worcestershire, gelatina de '
     'pescado'),
    ('Cacahuetes', 'Cacahuetes y derivados',
     'Salsas asiáticas, aceite de cacahuete, mantequilla de cacahuete, snacks, '
     'postres'),
    ('Soja', 'Semillas de soja y derivados',
     'Salsa de soja, tofu, lecitina (E322), aceite de soja, miso, tempeh, '
     'edamame'),
    ('Lácteos', 'Leche de vaca, cabra y oveja y sus derivados (incluida la '
     'lactosa)',
     'Quesos, nata, mantequilla, yogur, bechamel, helados, chocolate con '
     'leche'),
    ('Frutos de cáscara', 'Almendras, avellanas, nueces, anacardos, pecanas, '
     'pistachos, macadamias y nueces de Brasil',
     'Postres, pralinés, pesto, turrón, mazapán, aceites y mantequillas de '
     'frutos secos'),
    ('Apio', 'Apio y derivados, incluida la sal de apio',
     'Sopas, caldos, ensaladas, snacks, salsas, mezclas de especias'),
    ('Mostaza', 'Semillas de mostaza y derivados',
     'Salsas, vinagretas, encurtidos, marinados, especias, curry en polvo'),
    ('Sésamo', 'Semillas de sésamo y derivados',
     'Pan de hamburguesa, hummus, tahini, aceite de sésamo, sushi, ensaladas'),
    # TEC-32: el umbral sin la referencia analítica queda ambiguo frente a la
    # ficha técnica de un proveedor. El Anexo II del Reg. (UE) 1169/2011 lo
    # expresa como SO2.
    ('Sulfitos', 'Dióxido de azufre y sulfitos en concentración superior a '
     '10 mg/kg o 10 mg/L expresado como SO2',
     'Vino, cerveza, vinagre, frutas desecadas, crustáceos, patatas '
     'procesadas'),
    ('Altramuces', 'Altramuces y derivados',
     'Harina de altramuz en pan y bollería, snacks, sustitutos del café'),
    ('Moluscos', 'Mejillones, almejas, ostras, calamares, pulpo, caracoles y '
     'derivados',
     'Paella, fideuá, tapas, ensaladas de marisco, tinta de calamar'),
]

PROTOCOLO_ANTES = [
    'Escucha y ANOTA la alergia en la comanda con el nombre del alérgeno. '
    'Nunca la transmitas sólo de palabra.',
    'Consulta la matriz de alérgenos (registro 08) del plato COMPLETO: salsas, '
    'guarniciones, panes, fritos y aliños incluidos.',
    'Avisa al jefe de cocina antes de lanzar la comanda. El plato lo elabora '
    'UNA sola persona, de principio a fin.',
    'Cámbiate los guantes, lávate las manos y usa tabla, cuchillo y utensilios '
    'exclusivos en una zona de trabajo despejada.',
    'No compartas freidora, aceite, agua de cocción, plancha ni tostadora sin '
    'limpiarlas antes: es la vía de contaminación más frecuente.',
    'Si no puedes garantizar la ausencia del alérgeno, DILO: «no podemos '
    'garantizar que este plato esté libre de …», y ofrece una alternativa. '
    'Saca el plato aparte, tapado e identificado, y entrégalo EN MANO al '
    'comensal alérgico.',
]
# DOM-R2-17 (ronda 2): este cartel se cuelga en cocina para leerse BAJO PRESIÓN
# y el orden importa en minutos. La v2.0 ponía el 112 en el paso 9, detrás de
# «guarda el plato, el envase y la etiqueta» y de «no muevas a la persona»: el
# primer paso era de investigación y no de asistencia. En una anafilaxia se pide
# ayuda y se administra la adrenalina primero —a la vez—, y la conservación de
# la muestra es lo último.
PROTOCOLO_REACCION = [
    'Llama al 112 ante cualquier síntoma: dificultad para respirar, hinchazón '
    'de labios o lengua, mareo, vómitos o urticaria extensa. El 112 se llama '
    'por SÍNTOMAS, no por precaución. Es lo PRIMERO que se hace.',
    'Si la persona lleva adrenalina autoinyectable, ayúdala a usarla siguiendo '
    'sus indicaciones —se administra a la vez que se llama—. No la muevas ni '
    'la dejes sola.',
    'Deja de servir de inmediato y guarda el plato, el envase y la etiqueta '
    'del producto: harán falta para investigar qué pasó.',
    'Anota el incidente el mismo día en el registro 11 (Acciones Correctivas): '
    'plato, lote, quién lo elaboró y qué se hizo.',
]


def _post_14(wb, fname, cambios):
    ws = wb['14 Alérgenos']
    _lienzo(ws, 60, 6)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50

    cel = ws.cell(row=1, column=2,
                  value='LOS 14 ALÉRGENOS DE DECLARACIÓN OBLIGATORIA')
    cel.font = Font(bold=True, size=16)
    ws.merge_cells('B1:D1')
    ws.row_dimensions[1].height = 24
    cel = ws.cell(row=2, column=2,
                  value='Reglamento (UE) 1169/2011 — Real Decreto 126/2015')
    cel.font = Font(size=10, italic=True)
    ws.merge_cells('B2:D2')

    motor.cabecera(ws, 4, ['Nº', 'Alérgeno', 'Descripción',
                           'Dónde se encuentra con más frecuencia'],
                   [5, 25, 50, 50])
    for i, (nombre, desc, donde) in enumerate(ALERGENOS_14):
        fila = 5 + i
        for col, txt in ((1, i + 1), (2, nombre), (3, desc), (4, donde)):
            c = ws.cell(row=fila, column=col, value=txt)
            motor.calculada(c)
            if col >= 2:
                c.alignment = Alignment(horizontal='left', vertical='center',
                                        wrap_text=True)
            if col == 2:
                c.font = Font(bold=True, size=11)
        ws.row_dimensions[fila].height = _alto(max(desc, donde, key=len), 48,
                                               minimo=28)

    fila = 5 + len(ALERGENOS_14) + 1
    n_pasos = 0
    for titulo, pasos in (
            ('PROTOCOLO ANTE UNA ALERGIA DECLARADA — ANTES DE COCINAR Y AL '
             'SERVIR', PROTOCOLO_ANTES),
            ('SI UN COMENSAL EMPIEZA UNA REACCIÓN', PROTOCOLO_REACCION)):
        motor.banda(ws, fila, titulo, NC_14)
        ws.row_dimensions[fila].height = 20
        fila += 1
        for paso in pasos:
            n_pasos += 1
            c = ws.cell(row=fila, column=1, value=n_pasos)
            c.font = Font(bold=True, size=12)
            c.alignment = Alignment(horizontal='center', vertical='top')
            c.border = motor.BORDE
            t = ws.cell(row=fila, column=2, value=paso)
            t.font = Font(size=11)
            t.alignment = Alignment(horizontal='left', vertical='top',
                                    wrap_text=True)
            t.border = motor.BORDE
            for col in (3, 4):
                ws.cell(row=fila, column=col).border = motor.BORDE
            ws.merge_cells(start_row=fila, start_column=2,
                           end_row=fila, end_column=4)
            # TEC-09: una celda COMBINADA no desborda al vecino, corta duro en
            # el borde. Con wrap y altura calculada el paso completo llega al
            # papel.
            ws.row_dimensions[fila].height = _alto(paso, 118, minimo=22)
            fila += 1
        fila += 1

    motor.nota(ws, fila,
               'Imprime estas fichas y cuélgalas en la cocina y en la barra. '
               'Todo el personal, de cocina y de sala, tiene que conocer los '
               '14 alérgenos y este protocolo.', ncols=NC_14)
    motor.nota(ws, fila + 1,
               'La declaración de alérgenos de cada uno de tus platos va en el '
               'registro 08 (Matriz de alérgenos). Este cartel es la referencia '
               'y el protocolo; la matriz es el documento que se le enseña al '
               'cliente que lo pide.', ncols=NC_14)
    motor.nota(ws, fila + 2, motor.CONSERVACION, ncols=NC_14)
    motor.nota(ws, fila + 3, motor.MARCA, ncols=NC_14)
    motor.IMPRESION[(fname, ws.title)] = (4, True, False)
    cambios.append(f'14: protocolo de {n_pasos} pasos en dos bloques con wrap '
                   'y altura calculada, y sulfitos «expresado como SO2» '
                   '(DOM-27/TEC-09/TEC-32)')

    motor.escribir_instrucciones(wb, 'Fichas de los 14 Alérgenos Obligatorios', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Imprime las fichas y cuélgalas en la cocina y en la barra. Es un '
              'cartel de consulta, no un registro que se rellene.'),
        ('b', 'Obligatorio según el Reglamento (UE) 1169/2011 y el RD 126/2015: '
              'la información sobre los 14 alérgenos tiene que estar a '
              'disposición del cliente, y el personal tiene que saber darla.'),
        ('b', 'La declaración plato a plato de TU carta va en el registro 08 '
              '(Matriz de alérgenos). Este cartel es la referencia común y el '
              'protocolo de actuación.'),
        ('h', 'El protocolo, en dos bloques'),
        ('b', 'Antes de cocinar y al servir: seis pasos que evitan la '
              'contaminación cruzada, desde anotar la alergia en la comanda '
              'hasta entregar el plato en mano.'),
        ('b', 'Si hay una reacción: cuatro pasos en orden de urgencia. El '
              'PRIMERO es llamar al 112, y el segundo la adrenalina '
              'autoinyectable si la persona la lleva; guardar el plato y la '
              'etiqueta va después. Se llama por síntomas —dificultad para '
              'respirar, hinchazón de labios o lengua, mareo, vómitos, '
              'urticaria extensa—, no por precaución.'),
        ('b', 'Los dos bloques están pensados para leerse bajo presión: un paso '
              'por línea, numerados y sin letra pequeña.'),
        ('h', 'Lo que más falla en la práctica'),
        ('b', 'La freidora y el agua de cocción compartidas. Un rebozado con '
              'gluten deja gluten en el aceite; la pasta deja gluten en el '
              'agua.'),
        ('b', 'Los guantes. Se cambian ANTES de tocar el plato del comensal '
              'alérgico, no después.'),
        ('b', 'Decir «creo que no lleva». Si no está comprobado en la matriz, '
              'la respuesta honesta es que no se puede garantizar.'),
        ('b', motor.CONSERVACION),
    ], cambios)


# ==========================================================================
# 15 — Guía: cómo pasar una inspección de sanidad
# ==========================================================================
# DOM-10/COM-07: el producto anunciaba cuatro contenidos y entregaba uno, y esa
# única tabla tenía 23 puntos numerados mientras el título de la hoja, el nombre
# de la pestaña, la landing y el dashboard decían 25. Es la promesa más
# contable del pack: se verifica en diez segundos.
# DOM-28: la escala «GRAVE / MODERADA / LEVE» está inventada. En España las
# infracciones se clasifican como leves, graves y muy graves (Ley 17/2011,
# arts. 50-52), y el cierre cautelar es una medida excepcional del art. 27, no
# la consecuencia posible de cada punto marcado GRAVE.
# TEC-16: el resumen metía en el mismo saco un incumplimiento leve y uno grave.
NC_15 = 5
ESTADO_15 = '"✓ Cumple,⚠ Mejorar,✗ No cumple"'

PUNTOS_15 = [
    ('sec', 'DOCUMENTACIÓN Y REGISTROS'),
    ('p', 'Plan APPCC completo y actualizado',
     'Análisis de peligros, PCC, límites críticos, vigilancia, acciones '
     'correctivas y verificación (registro 12)', 'Grave'),
    ('p', 'Registros de control al día y firmados',
     'Temperaturas, limpieza, trazabilidad y acciones correctivas del último '
     'año, con firma', 'Grave'),
    ('p', 'Formación en higiene alimentaria acreditada por la empresa',
     'Justificante de formación de todos los manipuladores (Reg. (CE) '
     '852/2004, Anexo II, Cap. XII) y su registro (BONUS-01)', 'Grave'),
    ('p', 'Carta de alérgenos accesible y actualizada',
     'Disponible para cualquier cliente que la pida, coherente con la matriz '
     'del registro 08 y con la carta vigente', 'Muy grave'),
    ('p', 'Contrato de empresa DDD con Nº ROESB y certificados',
     'Contrato en vigor, empresa inscrita en el ROESB, certificados de cada '
     'actuación y plano de cebos (registro 07)', 'Grave'),
    ('sec', 'TEMPERATURAS Y EQUIPOS'),
    ('p', 'Cámaras de refrigeración (0-4 °C)',
     'Medición in situ con el termómetro del inspector, contrastada con tu '
     'registro 01', 'Grave'),
    ('p', 'Congeladores (≤ -18 °C)',
     'Medición in situ y producto correctamente congelado, sin escarcha ni '
     'recongelaciones', 'Grave'),
    ('p', 'Exposición caliente (≥ 65 °C)',
     'Temperatura de mantenimiento de los platos calientes en el momento de la '
     'visita', 'Grave'),
    ('p', 'Termómetro de sonda limpio, accesible y en uso',
     'Que exista, que esté a mano, que esté limpio y que el personal sepa '
     'usarlo', 'Leve'),
    ('p', 'Registro de verificación de termómetros',
     'Comprobación periódica de la sonda en hielo fundente o ebullición, con '
     'la desviación anotada (registro 19)', 'Grave'),
    ('sec', 'HIGIENE Y LIMPIEZA'),
    ('p', 'Limpieza general del local',
     'Suelos, paredes, techos y equipos sin grasa acumulada, incluida la parte '
     'de atrás de la maquinaria', 'Grave'),
    ('p', 'Estado de las superficies de trabajo',
     'Sin grietas ni desperfectos, de material apto para uso alimentario y '
     'fácil de limpiar', 'Leve'),
    ('p', 'Lavamanos en la zona de manipulación',
     'De accionamiento no manual, con agua caliente, jabón, papel de un solo '
     'uso y papelera de pedal', 'Grave'),
    ('p', 'Separación de zonas sucias y limpias',
     'Flujo de trabajo sin cruces entre crudo y cocinado, ni en el espacio ni '
     'en el tiempo', 'Grave'),
    ('p', 'Vestuarios del personal',
     'Ropa de calle separada de la de trabajo, taquillas y limpieza', 'Leve'),
    ('sec', 'ALMACENAMIENTO'),
    ('p', 'FIFO respetado, sin producto caducado',
     'Lo más antiguo delante; ningún caducado en cámara ni en almacén, tampoco '
     '«apartado para tirar»', 'Grave'),
    ('p', 'Separación de crudo y cocinado',
     'Nunca en el mismo estante; el crudo siempre por debajo del producto '
     'listo para consumo', 'Grave'),
    ('p', 'Etiquetado del producto elaborado',
     'Todo identificado con nombre, fecha de elaboración y fecha de consumo',
     'Leve'),
    ('p', 'Producto no apoyado en el suelo',
     'Todo sobre estanterías o palés, nada directamente en el suelo', 'Leve'),
    ('p', 'Productos de limpieza separados y en armario cerrado',
     'Nunca junto a alimentos, en su envase original y con su etiqueta y su '
     'ficha de datos de seguridad', 'Grave'),
    ('sec', 'MANIPULACIÓN Y SERVICIO'),
    ('p', 'Uniforme e higiene del personal',
     'Limpio, gorro, sin joyas, calzado cerrado, uñas cortas y heridas '
     'cubiertas', 'Leve'),
    # DOM-R2-16: freír por encima del límite de polares no es «un defecto
    # formal o de dotación sin riesgo directo» —que es como el propio pie de
    # esta hoja define «Leve»—: es servir alimento elaborado en un medio no
    # apto. Con «Leve» el resumen automático no lo contaba ni en muy graves ni
    # en graves, así que el punto salía en rojo y el resumen tranquilo.
    ('p', 'Aceite de fritura',
     'Test de compuestos polares, registro de cambios (09) y justificante de '
     'retirada por gestor autorizado', 'Grave'),
    ('p', 'Descongelación correcta',
     'En refrigeración a ≤4 °C, tapada y sobre bandeja; nunca a temperatura '
     'ambiente ni bajo el grifo', 'Grave'),
    ('p', 'Congelación preventiva de anisakis',
     'Pescado para consumo en crudo, marinado o escabechado: -20 °C 24 h o '
     '-35 °C 15 h, con registro (18)', 'Muy grave'),
    ('p', 'Control de alérgenos en servicio',
     'Personal formado, protocolo escrito y utensilios exclusivos ante una '
     'alergia declarada (ficha 14)', 'Muy grave'),
]

ANTES_15 = [
    'Rellena esta autoevaluación tú mismo y corrige lo que salga en rojo: el '
    'inspector va a mirar exactamente lo mismo.',
    'Pon al día los registros del mes en curso y fírmalos. Una hoja sin firmar '
    'cuenta como no hecha.',
    'Comprueba con sonda las temperaturas de cámaras, congeladores y baño '
    'maría, y anótalas en el registro 01.',
    'Revisa fechas de caducidad y el etiquetado de todo lo elaborado en casa.',
    'Ordena la cámara: crudos abajo, todo tapado, etiquetado y separado del '
    'producto listo para consumo.',
    'Repasa desagües, campana, juntas de puerta y la parte de atrás de los '
    'equipos: es donde se mira cuando lo evidente está limpio.',
    'Guarda los productos de limpieza en su armario cerrado y deja accesibles '
    'las fichas de datos de seguridad.',
    'Decide quién acompaña al inspector y quién sigue trabajando. Que sea '
    'alguien que sepa dónde está cada registro y que no improvise respuestas.',
]

DOCUMENTOS_15 = [
    'Plan APPCC: análisis de peligros (12) y plan de limpieza y desinfección '
    '(03) con su listado de productos químicos.',
    'Registros de temperaturas: diario (01) y de recepción (02) del último '
    'año.',
    'Checklist de recepción de mercancías (05) junto con sus albaranes.',
    'Registro de trazabilidad de entradas y salidas (06).',
    'Registro de limpieza diaria y de tareas semanales (04).',
    'Control de plagas: contrato con el Nº ROESB, certificados de cada '
    'actuación y plano de cebos (07).',
    'Control de aceite de fritura (09) y justificantes de retirada por gestor '
    'autorizado.',
    'Control de agua potable (10) y boletines de análisis si tienes depósito '
    'o pozo propio.',
    'Matriz de alérgenos (08) y carta de alérgenos a disposición del cliente.',
    'Registros de cocción y regeneración (16), enfriamiento y descongelación '
    '(17) y congelación preventiva de anisakis (18).',
    'Verificación de termómetros (19), registro de acciones correctivas (11) '
    'y registro de formación del personal (BONUS-01).',
]

ERRORES_15 = [
    'Registros rellenados a posteriori, todos con la misma letra y el mismo '
    'bolígrafo. Se nota, y anula el valor de la serie entera.',
    'Plan APPCC comprado y guardado en un cajón, sin una sola firma y sin una '
    'sola acción correctiva registrada en todo el año.',
    'Producto sin tapar, sin etiquetar o con crudo encima de cocinado en la '
    'cámara.',
    'Producto caducado en cámara o en almacén, aunque esté apartado «para '
    'tirar»: si está dentro, cuenta.',
    'Descongelar a temperatura ambiente o bajo el grifo.',
    'Servir pescado en crudo, marinado o escabechado sin congelación '
    'preventiva documentada.',
    'Decir «no lleva» un alérgeno sin haberlo comprobado en la matriz, o no '
    'tener carta de alérgenos.',
    'Lavamanos usado como fregadero, sin jabón, sin papel de un solo uso o sin '
    'papelera de pedal.',
    'Productos de limpieza junto a alimentos, o trasvasados a botellas sin '
    'etiquetar.',
]

N_PUNTOS_15 = sum(1 for t, *_ in PUNTOS_15 if t == 'p')
F0_15 = 5
F1_15 = F0_15 + len(PUNTOS_15) - 1              # última fila de la tabla (34)
FR_15 = F1_15 + 2                               # banda RESUMEN (36)


def _post_15(wb, fname, cambios):
    ws = wb['25 Puntos Inspección']
    _lienzo(ws, FR_15 + 60, NC_15 + 2)
    _titulo(ws, f'Los {N_PUNTOS_15} Puntos que Revisa el Inspector de Sanidad',
            NC_15,
            'Autoevalúa tu establecimiento ANTES de la inspección. Debajo de la '
            'tabla tienes el resumen automático y tres bloques: qué hacer las '
            '24 horas previas, qué documentos tener listos y los errores que '
            'más se sancionan.')
    motor.cabecera(ws, 4, ['Nº', 'Punto de inspección',
                           'Qué revisa el inspector',
                           'Gravedad (Ley 17/2011)', 'Tu estado'],
                   [6, 34, 56, 20, 18])

    fila, n = F0_15, 0
    filas_dato = []
    for item in PUNTOS_15:
        if item[0] == 'sec':
            motor.banda(ws, fila, item[1], NC_15)
            ws.row_dimensions[fila].height = 20
            fila += 1
            continue
        _, punto, revisa, gravedad = item
        n += 1
        filas_dato.append(fila)
        c = ws.cell(row=fila, column=1, value=n)
        motor.calculada(c, motor.FMT_ENT)
        for col, txt in ((2, punto), (3, revisa)):
            c = ws.cell(row=fila, column=col, value=txt)
            motor.calculada(c)
            c.alignment = Alignment(horizontal='left', vertical='center',
                                    wrap_text=True)
        c = ws.cell(row=fila, column=4, value=gravedad)
        motor.calculada(c)
        c.font = Font(bold=True, size=11)
        motor.verde(ws.cell(row=fila, column=5), '@')
        ws.row_dimensions[fila].height = _alto(max(punto, revisa, key=len), 54,
                                               minimo=28)
        fila += 1

    for r in _rangos_filas('E', filas_dato):
        motor.dv_lista(ws, ESTADO_15, [r], 'Tu estado',
                       'Elige «✓ Cumple», «⚠ Mejorar» o «✗ No cumple». El '
                       'resumen compara por igualdad exacta, así que un texto '
                       'escrito a mano no suma.')
        motor.semaforo(ws, r, extra_ok=('✓ Cumple',),
                       extra_ambar=('⚠ Mejorar',),
                       extra_rojo=('✗ No cumple',))
    for r in _rangos_filas('D', filas_dato):
        motor.semaforo(ws, r, extra_ok=('Leve',), extra_ambar=('Grave',),
                       extra_rojo=('Muy grave',))

    # TEC-16: el dato accionable —cuántos puntos GRAVES o MUY GRAVES has
    # suspendido— era justo el que no se calculaba, aunque la columna de
    # gravedad ya estaba ahí. Un local con 3 graves fallados y 2 leves leía
    # «5 puntos que no cumplen» y no distinguía el riesgo del detalle menor.
    rd = f'$D${F0_15}:$D${F1_15}'
    re_ = f'$E${F0_15}:$E${F1_15}'
    motor.banda(ws, FR_15, 'RESUMEN DE LA AUTOEVALUACIÓN', NC_15)
    resp = f'E{FR_15 + 1}+E{FR_15 + 2}+E{FR_15 + 3}'
    filas_res = [
        ('Puntos que cumplen', f'=COUNTIF({re_},"✓ Cumple")', None, None),
        ('Puntos a mejorar', f'=COUNTIF({re_},"⚠ Mejorar")', None, None),
        ('Puntos que no cumplen', f'=COUNTIF({re_},"✗ No cumple")', None, None),
        # DOM-R2-01 / TEC-02 (altas, ronda 2): faltaba el paréntesis.
        # «=25-E37+E38+E39» es 25 − E37 + E38 + E39: sólo acertaba con la hoja
        # vacía o con los 25 puntos en «✓ Cumple», y con una mezcla de los tres
        # estados llegaba a declarar 50 puntos sin responder sobre 25 — con la
        # celda en rojo permanente por su propia regla «=E40>0». MAX(0;…) por si
        # alguien amplía la lista de puntos sin tocar la constante.
        (f'Sin responder (de {N_PUNTOS_15})',
         f'=MAX(0,{N_PUNTOS_15}-({resp}))', None, f'=E{FR_15 + 4}>0'),
        ('Incumplimientos MUY GRAVES (riesgo de medidas cautelares)',
         f'=COUNTIFS({rd},"Muy grave",{re_},"✗ No cumple")', None,
         f'=E{FR_15 + 5}>0'),
        ('Incumplimientos GRAVES',
         f'=COUNTIFS({rd},"Grave",{re_},"✗ No cumple")', None,
         f'=E{FR_15 + 6}>0'),
        ('% de cumplimiento (sobre los puntos respondidos)',
         f'=IF(({resp})=0,"",E{FR_15 + 1}/({resp}))', '0%', None),
    ]
    for i, (etiqueta, formula, fmt, rojo) in enumerate(filas_res):
        _contador(ws, FR_15 + 1 + i, NC_15, 5, etiqueta, formula, fmt, rojo)

    fila = FR_15 + len(filas_res) + 2
    for titulo, bloque in (
            ('ANTES DE LA INSPECCIÓN — LAS 24 HORAS PREVIAS', ANTES_15),
            ('DOCUMENTOS QUE TENER LISTOS, EN UNA CARPETA Y A MANO',
             DOCUMENTOS_15),
            ('LOS ERRORES QUE MÁS SE SANCIONAN', ERRORES_15)):
        motor.banda(ws, fila, titulo, NC_15)
        ws.row_dimensions[fila].height = 20
        fila += 1
        for i, texto in enumerate(bloque, start=1):
            c = ws.cell(row=fila, column=1, value=i)
            c.font = Font(bold=True, size=11)
            c.alignment = Alignment(horizontal='center', vertical='top')
            t = ws.cell(row=fila, column=2, value=texto)
            t.font = Font(size=11)
            t.alignment = Alignment(horizontal='left', vertical='top',
                                    wrap_text=True)
            ws.merge_cells(start_row=fila, start_column=2,
                           end_row=fila, end_column=NC_15)
            ws.row_dimensions[fila].height = _alto(texto, 122, minimo=20)
            fila += 1
        fila += 1

    _pie(ws, fila, NC_15, extra=(
        # COM-R2-12: el art. 50 tipifica las TRES categorías y el 51 fija las
        # sanciones; repartir una categoría por artículo es lo que un inspector
        # o un consultor detecta al primer vistazo. El 12 ya lo cita bien, como
        # rango (arts. 50-52), y así se queda también aquí.
        'Escala de gravedad según la Ley 17/2011, de seguridad alimentaria y '
        'nutrición (arts. 50 a 52): leve, grave y muy grave. '
        'Muy grave = el incumplimiento pone en el mercado alimentos con riesgo '
        'para la salud; grave = incumplimiento de una obligación de '
        'autocontrol, higiene o registro; leve = defecto formal o de dotación '
        'sin riesgo directo.',
        'Los incumplimientos calificados como muy graves, o cualquier '
        'situación de riesgo grave e inminente para la salud, pueden dar lugar '
        'a medidas cautelares incluido el cierre del establecimiento '
        '(Ley 17/2011, art. 27). Las sanciones van ' + motor.SANCIONES + '.',
        'Esta guía es una autoevaluación de preparación, no una resolución '
        'administrativa: la calificación de una infracción la hace siempre la '
        'autoridad competente, valorando el riesgo, la intencionalidad y la '
        'reincidencia.'))
    motor.IMPRESION[(fname, ws.title)] = (4, True)
    cambios.append(f'15: {n} puntos reales (añadidos verificación de '
                   'termómetros y congelación preventiva de anisakis), escala '
                   'de la Ley 17/2011, resumen con COUNTIFS de muy graves y '
                   'graves, % de cumplimiento y «Sin responder», y bloques de '
                   'preparación, documentos y errores '
                   '(DOM-10/DOM-28/TEC-16/COM-07)')

    motor.escribir_instrucciones(
        wb, 'Guía: Cómo Pasar una Inspección de Sanidad', [
            ('h', 'Qué encontrarás en esta guía'),
            ('b', f'Los {N_PUNTOS_15} puntos que revisa el inspector, con lo '
                  'que mira en cada uno y su gravedad según la Ley 17/2011.'),
            ('b', 'Un resumen automático que separa los incumplimientos muy '
                  'graves de los graves, cuenta los puntos sin responder y '
                  'calcula tu porcentaje de cumplimiento.'),
            ('b', 'Qué hacer en las 24 horas previas a una inspección.'),
            ('b', 'Qué documentos tener listos, en una carpeta y a mano.'),
            ('b', 'Los errores que más se sancionan, para que no te pillen en '
                  'ninguno.'),
            # DOM-R2-23 (ronda 2) pedía cambiar esta línea por «29 años en
            # alta hostelería y 15 como consultor gastronómico». NO se aplica,
            # y no por criterio sino por gate: `censo-entregables.py:139`
            # (`RX_BIO_VIEJA`) clasifica «29 años» y «15 años» como la bio
            # ANTIGUA y tumba el censo con `bio_vieja` — comprobado aplicando
            # el cambio: «FAIL: 15-guia-inspeccion-sanidad.xlsx: bio_vieja».
            # La formulación de abajo es la que dejó el saneamiento de bios del
            # 2026-08-18/22 y la que llevan los otros 43 productos
            # (p. ej. `kit-tareas-bar`, `plan-negocio-bar-restaurante`):
            # cambiarla sólo aquí desincronizaría este fichero del catálogo.
            ('b', 'Basada en la experiencia de John Guerrero: consultor '
                  'gastronómico desde 2010, en cocina desde los 17 años.'),
            ('h', 'Cómo se usa'),
            ('b', 'Recorre el local con la hoja impresa o en el móvil y marca '
                  'cada punto con el desplegable: ✓ Cumple, ⚠ Mejorar o '
                  '✗ No cumple. Se pintan solos.'),
            ('b', 'Mira primero la casilla de «Incumplimientos MUY GRAVES»: si '
                  'no está en cero, eso es lo que hay que resolver hoy.'),
            ('b', 'Repite la autoevaluación cada trimestre y guarda las hojas. '
                  'La serie demuestra que revisas tu propio sistema, que es '
                  'justo lo que pide el principio 6 del APPCC.'),
            ('h', 'Sobre la escala de gravedad'),
            ('p', 'En España las infracciones en seguridad alimentaria se '
                  'clasifican en leves, graves y muy graves (Ley 17/2011, '
                  'arts. 50-52). El cierre cautelar no es la consecuencia '
                  'normal de un incumplimiento: es una medida excepcional '
                  'reservada a las situaciones de riesgo grave e inminente '
                  'para la salud (art. 27). La calificación concreta la hace '
                  'siempre la autoridad competente.'),
            ('b', motor.CONSERVACION),
        ], cambios)


# ==========================================================================
# API del grupo
# ==========================================================================
POST = {
    '03-plan-limpieza-desinfeccion.xlsx': _post_03,
    '04-registro-limpieza-diaria.xlsx': _post_04,
    '07-control-plagas-ddd.xlsx': _post_07,
    '11-registro-acciones-correctivas.xlsx': _post_11,
    '12-analisis-peligros-haccp.xlsx': _post_12,
    '13-checklist-higiene-personal.xlsx': _post_13,
    '14-fichas-14-alergenos.xlsx': _post_14,
    '15-guia-inspeccion-sanidad.xlsx': _post_15,
}


def pre(wb, fname, cambios):
    """Ninguna inserción estructural: cada hoja se reconstruye entera en
    `post()`, así que no hay columnas que desplazar ni centinelas que vigilar."""
    return


def post(wb, fname, cambios, registro=None):
    fn = POST.get(fname)
    if fn:
        fn(wb, fname, cambios)


# ==========================================================================
# Casos «dato FUERA de límite» que consume `main.py` (paso 5)
# ==========================================================================
# En el grupo A esto demostraba que una temperatura mala deja de dar OK. Aquí
# los veredictos son de otra naturaleza —nivel de riesgo, recuentos de
# incumplimientos, tareas pendientes—, pero la exigencia es la misma: meter un
# dato malo y ver que el contador se mueve. Un contador que devolviera siempre
# cero pasaría cualquier revisión visual.
_F_LIBRE_03 = F_DATOS_03[-1]                    # última fila libre del 03
_F_LIBRE_12A = F1_12 + 2                        # filas libres del 12
_F_LIBRE_12B = F1_12 + 3
_FILAS_PUNTO_15 = []                            # las 25 filas de respuesta
_i = F0_15
for _item in PUNTOS_15:
    if _item[0] == 'sec':
        _i += 1
        continue
    _FILAS_PUNTO_15.append(_i)
    _i += 1
_F_PUNTO23_15 = _FILAS_PUNTO_15[22]
_F_PUNTO24_15 = _FILAS_PUNTO_15[23]
assert len(_FILAS_PUNTO_15) == N_PUNTOS_15

CASOS_LIMITE = [
    {'fichero': '12-analisis-peligros-haccp.xlsx', 'hoja': 'Análisis Peligros',
     'entradas': {f'E{_F_LIBRE_12A}': 'Alta', f'F{_F_LIBRE_12A}': 'Alta'},
     'salida': f'G{_F_LIBRE_12A}', 'esperado': 'Crítico',
     'lectura': 'Peligro con probabilidad Alta y gravedad Alta → nivel de '
                'riesgo Crítico (en v1.1 las dos columnas no alimentaban '
                'ninguna fórmula)'},
    {'fichero': '12-analisis-peligros-haccp.xlsx', 'hoja': 'Análisis Peligros',
     'entradas': {f'E{_F_LIBRE_12B}': 'Media', f'F{_F_LIBRE_12B}': 'Alta'},
     'salida': f'G{_F_LIBRE_12B}', 'esperado': 'Alto',
     'lectura': 'Probabilidad Media y gravedad Alta → Alto. Es la puntuación '
                'que en v1.1 se resolvía seis veces como PCC y dos como NO'},
    {'fichero': '15-guia-inspeccion-sanidad.xlsx', 'hoja': '25 Puntos Inspección',
     'entradas': {f'E{_F_PUNTO24_15}': '✗ No cumple'},
     'salida': f'E{FR_15 + 5}', 'esperado': 1,
     'lectura': 'Congelación preventiva de anisakis incumplida → 1 '
                'incumplimiento MUY GRAVE (TEC-16: el resumen de v1.1 no '
                'distinguía la gravedad)'},
    {'fichero': '15-guia-inspeccion-sanidad.xlsx', 'hoja': '25 Puntos Inspección',
     'entradas': {f'E{_F_PUNTO23_15}': '✗ No cumple'},
     'salida': f'E{FR_15 + 6}', 'esperado': 1,
     'lectura': 'Descongelación incorrecta (Grave) → 1 incumplimiento GRAVE y '
                'CERO muy graves: «Grave» no casa con «Muy grave» en el '
                'COUNTIFS'},
    {'fichero': '15-guia-inspeccion-sanidad.xlsx', 'hoja': '25 Puntos Inspección',
     'entradas': {f'E{_F_PUNTO24_15}': '✓ Cumple'},
     'salida': f'E{FR_15 + 4}', 'esperado': N_PUNTOS_15 - 1,
     'lectura': f'Un punto respondido → «Sin responder» baja a '
                f'{N_PUNTOS_15 - 1}'},
    # DOM-R2-01 / TEC-02: el hueco por el que se coló la fórmula sin
    # paréntesis. Los dos casos que había —hoja vacía y un solo «✓ Cumple»—
    # daban verde con «=25-E37+E38+E39». Con los 25 puntos contestados y una
    # MEZCLA de los tres estados, la fórmula rota declaraba 4, 10 o hasta 50
    # puntos sin responder; la buena dice 0.
    {'fichero': '15-guia-inspeccion-sanidad.xlsx', 'hoja': '25 Puntos Inspección',
     'entradas': dict(
         [(f'E{f}', '✓ Cumple') for f in _FILAS_PUNTO_15[:20]]
         + [(f'E{f}', '⚠ Mejorar') for f in _FILAS_PUNTO_15[20:23]]
         + [(f'E{f}', '✗ No cumple') for f in _FILAS_PUNTO_15[23:]]),
     'salida': f'E{FR_15 + 4}', 'esperado': 0,
     'lectura': 'Los 25 puntos contestados con una mezcla de ✓ Cumple, '
                '⚠ Mejorar y ✗ No cumple → «Sin responder» = 0. Con la '
                'fórmula sin paréntesis daba 10 (DOM-R2-01/TEC-02)'},
    {'fichero': '15-guia-inspeccion-sanidad.xlsx', 'hoja': '25 Puntos Inspección',
     'entradas': dict([(f'E{f}', '✗ No cumple') for f in _FILAS_PUNTO_15]),
     'salida': f'E{FR_15 + 4}', 'esperado': 0,
     'lectura': 'Los 25 puntos en «✗ No cumple» → «Sin responder» = 0. La '
                'fórmula rota decía 50 sobre 25'},
    {'fichero': '03-plan-limpieza-desinfeccion.xlsx', 'hoja': 'Plan Maestro L+D',
     'entradas': {f'F{_F_LIBRE_03}': 'N'}, 'salida': f'F{F_CONT_03}',
     # 2 y no 1: la línea de fregaderos se entrega ya con «N» para que el
     # ejemplo case con el desincrustante ácido de la pestaña «Productos
     # químicos», que llega con la FDS pendiente (DOM-R2-12).
     'esperado': 2,
     'lectura': 'Una línea más con producto químico sin nº de registro ni '
                'ficha de datos de seguridad → el contador pasa de 1 a 2 '
                '(DOM-12/DOM-R2-12)'},
    {'fichero': '04-registro-limpieza-diaria.xlsx', 'hoja': 'Limpieza Diaria',
     'entradas': {f'D{FILAS_TAREA_04[5]}': '✗'}, 'salida': f'B{F_CONT_04}',
     'esperado': 1,
     'lectura': 'Una tarea marcada como no realizada el martes por la mañana → '
                'el contador semanal la suma'},
    {'fichero': '07-control-plagas-ddd.xlsx', 'hoja': HOJA_CEBOS,
     'entradas': {f'F{F0_07C + 2}': 'Consumo total'},
     'salida': f'F{F_CONT_07C}', 'esperado': 1,
     'lectura': 'Cebadero con consumo total → estación con actividad (DOM-16)'},
    {'fichero': '11-registro-acciones-correctivas.xlsx',
     'hoja': 'Acciones Correctivas',
     'entradas': {f'M{F0_11 + 10}': 'Pendiente'}, 'salida': f'M{F_CONT_11}',
     'esperado': 2,
     'lectura': 'Una incidencia más sin verificar → el contador de abiertas '
                'pasa de 1 a 2 (la hoja se entrega con INC-003 en «Requiere '
                'seguimiento», a propósito: un registro de acciones '
                'correctivas con todo cerrado desde el primer día no se lo '
                'cree ningún inspector)'},
]


# ==========================================================================
# §6 — lo que hay que poder demostrar en la ronda de refutación
# ==========================================================================
def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}: {e}'


def _caso(carpeta, destino, fname, hoja, entradas, salida, esperado, lectura):
    """Escribe `entradas` en una COPIA desechable y evalúa `salida` con pycel."""
    from pycel import ExcelCompiler
    import openpyxl
    os.makedirs(destino, exist_ok=True)
    etiqueta = f"{fname.split('-')[0]}-{hoja[:6].replace(' ', '')}-{salida}"
    dst = os.path.join(destino, f'demoB-{etiqueta}.xlsx')
    shutil.copy2(os.path.join(carpeta, fname), dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb[hoja]
    for coord, valor in entradas.items():
        ws[coord] = valor
    wb.save(dst)
    obtenido = _ev(ExcelCompiler(filename=dst), f"'{hoja}'!{salida}")
    return {
        'ref': f'{fname}:{hoja}:{salida}',
        'entradas': {f'{hoja}!{k}': v for k, v in entradas.items()},
        'esperado': esperado, 'obtenido': obtenido,
        'ok': obtenido == esperado, 'lectura': lectura,
        'copia_desechable': dst,
    }


ACIDOS = ('desincrustante', 'ácido', 'antical', 'descalcificador')
CLORO = ('lejía', 'hipoclorito')


def _conteos(carpeta):
    """Los recuentos que la SPEC pide poder demostrar, leídos del fichero YA
    generado, no de las constantes de este módulo: si un bucle se saltara una
    fila, la constante seguiría diciendo 25 y el xlsx tendría 24."""
    import openpyxl
    fuera, fallos = {}, []

    # 15 — «25 puntos contados»
    p = os.path.join(carpeta, '15-guia-inspeccion-sanidad.xlsx')
    if os.path.isfile(p):
        ws = openpyxl.load_workbook(p)['25 Puntos Inspección']
        nums = [ws.cell(row=r, column=1).value
                for r in range(F0_15, F1_15 + 1)
                if isinstance(ws.cell(row=r, column=1).value, int)]
        fuera['15_puntos_numerados'] = len(nums)
        fuera['15_numeracion_correlativa'] = nums == list(range(1, len(nums) + 1))
        titulo = str(ws['A1'].value or '')
        fuera['15_titulo'] = titulo
        if len(nums) != 25:
            fallos.append(f'§6 15: la tabla tiene {len(nums)} puntos '
                          'numerados, no 25')
        if '25' not in titulo:
            fallos.append(f'§6 15: el título ya no dice 25 → {titulo!r}')
        gravedades = {ws.cell(row=r, column=4).value
                      for r in range(F0_15, F1_15 + 1)
                      if ws.cell(row=r, column=4).value}
        fuera['15_escala_gravedad'] = sorted(gravedades)
        if gravedades - {'Leve', 'Grave', 'Muy grave'}:
            fallos.append(f'§6 15: escala fuera de la Ley 17/2011 → '
                          f'{sorted(gravedades)}')

    # 12 — «7 fases»
    p = os.path.join(carpeta, '12-analisis-peligros-haccp.xlsx')
    if os.path.isfile(p):
        ws = openpyxl.load_workbook(p)['Análisis Peligros']
        fases = []
        for r in range(F0_12, FL1_12 + 1):
            v = ws.cell(row=r, column=1).value
            if v and v not in fases:
                fases.append(v)
        fuera['12_fases'] = fases
        fuera['12_peligros'] = sum(
            1 for r in range(F0_12, FL1_12 + 1)
            if ws.cell(row=r, column=3).value)
        if len(fases) != 7:
            fallos.append(f'§6 12: {len(fases)} fases, no 7 → {fases}')
        # DOM-01/TEC-04: la columna «Registro» sólo puede citar entregables.
        entregables = ('01', '02', '03', '04', '05', '06', '07', '08', '09',
                       '10', '11', '12', '13', '14', '15', '16', '17', '18',
                       '19', 'BONUS')
        malos = []
        for r in range(F0_12, FL1_12 + 1):
            v = ws.cell(row=r, column=10).value
            if v and not str(v).strip().startswith(entregables):
                malos.append(f'{ws.title}!J{r}={v!r}')
        fuera['12_registros_inexistentes'] = malos
        if malos:
            fallos.append('§6 12: la columna «Registro» cita algo que no es un '
                          'entregable del pack: ' + '; '.join(malos))
        # Coherencia riesgo ↔ decisión (TEC-08): misma puntuación, misma
        # respuesta. Se comprueba que no haya dos filas con idéntico par
        # (Probabilidad, Gravedad) resueltas una como PCC y otra como NO.
        combos = {}
        for r in range(F0_12, F1_12 + 1):
            clave = (ws.cell(row=r, column=5).value,
                     ws.cell(row=r, column=6).value)
            combos.setdefault(clave, set()).add(ws.cell(row=r, column=8).value)
        incoherentes = {str(k): sorted(v) for k, v in combos.items()
                        if {'PCC', 'NO'} <= v}
        fuera['12_puntuaciones_incoherentes'] = incoherentes
        if incoherentes:
            fallos.append('§6 12: la misma puntuación se resuelve como PCC y '
                          f'como NO → {incoherentes}')

    # 03 — fregaderos sin «lejía» junto a «desincrustante»
    p = os.path.join(carpeta, '03-plan-limpieza-desinfeccion.xlsx')
    if os.path.isfile(p):
        wb = openpyxl.load_workbook(p)
        ws = wb['Plan Maestro L+D']
        mezclas = []
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str):
                    continue
                t = c.value.lower()
                if any(a in t for a in ACIDOS) and any(x in t for x in CLORO):
                    # Sólo es defecto si NO separa las dos pasadas.
                    if not ('segunda pasada' in t or 'por separado' in t
                            or 'nunca mezclar' in t):
                        mezclas.append(f'{ws.title}!{c.coordinate}: '
                                       f'{c.value[:90]!r}')
        fuera['03_celdas_que_mezclan_acido_y_cloro'] = mezclas
        fila_freg = next((r for r in F_DATOS_03
                          if str(ws.cell(row=r, column=1).value or '')
                          .lower().startswith('fregadero')), None)
        prod = str(ws.cell(row=fila_freg, column=4).value or '') if fila_freg \
            else ''
        fuera['03_fregaderos_producto'] = prod
        fuera['03_fregaderos_sin_lejia'] = 'lejía' not in prod.lower()
        fuera['03_elementos'] = sum(
            1 for r in F_DATOS_03 if ws.cell(row=r, column=1).value)
        fuera['03_zonas'] = [ws.cell(row=r, column=1).value
                             for r in F_BANDAS_03]
        if mezclas:
            fallos.append('§6 03: hay celdas que siguen prescribiendo ácido y '
                          'cloro juntos sin separarlos: ' + '; '.join(mezclas))
        if fila_freg is None:
            fallos.append('§6 03: no encuentro la fila de fregaderos')
        elif 'lejía' in prod.lower() and 'desincrustante' in prod.lower():
            fallos.append(f'§6 03: los fregaderos siguen con desincrustante y '
                          f'lejía en la misma celda → {prod!r}')
        if fuera['03_elementos'] < 26:
            fallos.append(f"§6 03: {fuera['03_elementos']} elementos, la SPEC "
                          'pide ≥ 26')

    # 14 — pasos del protocolo
    p = os.path.join(carpeta, '14-fichas-14-alergenos.xlsx')
    if os.path.isfile(p):
        ws = openpyxl.load_workbook(p)['14 Alérgenos']
        # Sólo DEBAJO de la tabla de alérgenos: la columna A numera primero los
        # 14 alérgenos y después los pasos del protocolo, y contar la columna
        # entera daba 24 «pasos».
        pasos = [ws.cell(row=r, column=1).value
                 for r in range(5 + len(ALERGENOS_14) + 1, ws.max_row + 1)
                 if isinstance(ws.cell(row=r, column=1).value, int)]
        fuera['14_pasos_protocolo'] = len(pasos)
        fuera['14_menciona_112'] = any(
            isinstance(c.value, str) and '112' in c.value
            for row in ws.iter_rows() for c in row)
        if not 8 <= len(pasos) <= 10:
            fallos.append(f'§6 14: el protocolo tiene {len(pasos)} pasos, la '
                          'SPEC pide 8-10')
        if not fuera['14_menciona_112']:
            fallos.append('§6 14: el protocolo no menciona el 112')

    # 04 — la DV de marcas cubre B6:H28 (literal de la SPEC) y hasta la O
    p = os.path.join(carpeta, '04-registro-limpieza-diaria.xlsx')
    if os.path.isfile(p):
        ws = openpyxl.load_workbook(p)['Limpieza Diaria']
        cubiertas = set()
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list' and '✓' in str(dv.formula1):
                for rango in dv.sqref.ranges:
                    for fila in rango.rows:
                        for coord in fila:
                            cubiertas.add((coord[0], coord[1]))
        faltan = [f'{get_column_letter(c)}{r}'
                  for r in FILAS_TAREA_04 for c in range(2, 9)
                  if (r, c) not in cubiertas]
        fuera['04_dv_marcas_B6H28_completa'] = not faltan
        fuera['04_columnas_por_dia'] = 2
        if faltan:
            fallos.append('§6 04: sin desplegable de marcas en ' +
                          ', '.join(faltan[:8]))

    # 07 — filas y pestaña de cebos
    p = os.path.join(carpeta, '07-control-plagas-ddd.xlsx')
    if os.path.isfile(p):
        wb = openpyxl.load_workbook(p)
        fuera['07_filas_registro'] = F1_07 - F0_07 + 1
        fuera['07_hojas'] = wb.sheetnames
        cab = str(wb['Control Plagas DDD']['A2'].value or '')
        fuera['07_roesb_en_cabecera'] = 'ROESB' in cab
        if HOJA_CEBOS not in wb.sheetnames:
            fallos.append('§6 07: falta la pestaña «Plano de cebos»')
        if 'ROESB' not in cab:
            fallos.append('§6 07: el Nº ROESB no está en la cabecera')

    return fuera, fallos


def demos(carpeta, origen, destino):
    """Los casos del §6 que tocan al grupo B, los recuentos que la SPEC pide
    poder contar y el barrido de normas derogadas."""
    import openpyxl
    casos = []

    casos.append(_caso(
        carpeta, destino, '12-analisis-peligros-haccp.xlsx',
        'Análisis Peligros',
        {f'E{_F_LIBRE_12A}': 'Baja', f'F{_F_LIBRE_12A}': 'Baja'},
        f'G{_F_LIBRE_12A}', 'Bajo',
        'Probabilidad Baja y gravedad Baja → Bajo. El nivel de riesgo se '
        'calcula; en v1.1 las columnas Probabilidad y Gravedad no alimentaban '
        'ninguna fórmula (TEC-08).'))
    casos.append(_caso(
        carpeta, destino, '12-analisis-peligros-haccp.xlsx',
        'Análisis Peligros', {}, f'G{F0_12}', 'Crítico',
        'La primera fila entregada (recepción, Alta/Alta) ya viene con su '
        'nivel de riesgo resuelto: Crítico, coherente con su decisión de PCC.'))
    casos.append(_caso(
        carpeta, destino, '15-guia-inspeccion-sanidad.xlsx',
        '25 Puntos Inspección', {}, f'E{FR_15 + 4}', N_PUNTOS_15,
        f'Hoja recién entregada: los {N_PUNTOS_15} puntos figuran como «Sin '
        'responder», no como cumplidos. En v1.1 no existía ese contador y una '
        'hoja en blanco parecía una hoja aprobada (TEC-16).'))
    casos.append(_caso(
        carpeta, destino, '15-guia-inspeccion-sanidad.xlsx',
        '25 Puntos Inspección',
        {f'E{_F_PUNTO24_15}': '✗ No cumple', f'E{_F_PUNTO23_15}': '✗ No cumple'},
        f'E{FR_15 + 5}', 1,
        'Un incumplimiento muy grave y uno grave a la vez → el contador de MUY '
        'GRAVES dice 1, no 2: el COUNTIFS distingue «Grave» de «Muy grave».'))
    casos.append(_caso(
        carpeta, destino, '15-guia-inspeccion-sanidad.xlsx',
        '25 Puntos Inspección',
        {f'E{_F_PUNTO24_15}': '✗ No cumple', f'E{_F_PUNTO23_15}': '✗ No cumple'},
        f'E{FR_15 + 6}', 1,
        'El mismo caso leído por el otro contador: 1 incumplimiento GRAVE.'))

    conteos, fallos_conteo = _conteos(carpeta)

    # §6: «RD 2207/1995», «RD 140/2003», «carné de manipulador» y «€60.000»
    # con 0 ocurrencias en los ficheros del grupo.
    normativa = []
    for fname in FICHEROS:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for etiqueta, rx in motor.PROHIBIDAS:
            donde = [f'{ws.title}!{c.coordinate}'
                     for ws in wb.worksheets for row in ws.iter_rows()
                     for c in row
                     if isinstance(c.value, str) and rx.search(c.value)]
            normativa.append({'fichero': fname, 'cadena': etiqueta,
                              'patron': rx.pattern,
                              'ocurrencias': len(donde), 'donde': donde})

    fallos = [f"§6 {c['ref']}: esperaba {c['esperado']!r}, dio {c['obtenido']!r}"
              for c in casos if not c['ok']]
    fallos += fallos_conteo
    fallos += [f"§6 «{n['cadena']}» sigue viva en {n['fichero']}: {n['donde']}"
               for n in normativa if n['ocurrencias']]

    return {'fallos': fallos,
            'casos_spec_6': casos,
            'conteos_spec_6': conteos,
            'normativa_derogada': normativa,
            'normativa_ocurrencias_totales':
                sum(n['ocurrencias'] for n in normativa)}
