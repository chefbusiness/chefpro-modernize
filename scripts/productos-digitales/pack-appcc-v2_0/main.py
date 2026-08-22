#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 del Pack de Plantillas APPCC.

    python3 main.py --dry-run [--solo a|b|c|a,b] [--json informe.json]

REGLA DURA: `astro-site/public/dl/pack-appcc/` NO se toca. `--dry-run`
regenera una copia en el scratchpad y trabaja allí. Sin `--dry-run` el script
ABORTA salvo que se le pase `PACK_APPCC_APPLY=1` en el entorno: la ejecución
real la hace el orquestador cuando la ronda 2 dé verde
(`pack-appcc-v2-SPEC.md`, cabecera).

Qué hace, en este orden:
  1. Copia de trabajo (dry-run) desde `dl/pack-appcc`. En producción, respaldo
     con marca de tiempo en el SCRATCHPAD (nunca junto a los entregables: una
     carpeta `.bak` dentro de `public/dl/` se publicaría).
  2. Por fichero: grupo.pre → motor.aplicar (§1) → grupo.post (§2-§4) →
     motor.cerrar.
  3. Idempotencia: repite el pipeline sobre un clon del resultado y compara
     celda a celda. Debe dar 0 diferencias.
  4. `inject_cache.py` sobre los ficheros tocados (SIEMPRE al final: cualquier
     guardado posterior de openpyxl borraría el cache).
  5. Verificación `data_only` de todas las fórmulas nuevas.
  6. pycel con inputs FUERA de límite (§6): cada fórmula de Estado tiene que
     pasar de OK a ALERTA / RECHAZAR / CAMBIAR / REVISAR con un dato malo.
  6bis. Gate sobre el XML CRUDO (TEC-09): ningún `<formula>` de `cfRule` que
     empiece por «=», ningún `<mergeCell ref>` de una sola celda y ningún
     `<v></v>` vacío en celda de fórmula. Toda la demás verificación relee con
     la misma openpyxl que escribió, así que es ciega a lo que Excel repara.
  7. `censo-entregables.py --only <carpeta> --fail`.
  8. Demostraciones de cada grupo (§6) sobre copias desechables, nunca sobre
     los entregables, y barrido de normas derogadas.

Térmica: todo en SERIE. No hay builds ni navegador.
"""
import argparse
import contextlib
import datetime
import importlib
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
ORIGEN = os.path.join(ROOT, 'astro-site', 'public', 'dl', 'pack-appcc')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    'f83a04d3-d14f-49b3-aa79-c442fb4d7983/scratchpad')
DESTINO = os.path.join(SCRATCH, 'dryrun-appcc', 'pack-appcc')
IDEM = os.path.join(SCRATCH, 'dryrun-appcc', 'pack-appcc-idem')
DEMOS = os.path.join(SCRATCH, 'dryrun-appcc', '_demos')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')

SPEC = 'scripts/productos-digitales/pack-appcc-v2-SPEC.md §1 y §2'


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Copia de trabajo
# ==========================================================================
def preparar_copia():
    if not os.path.isdir(ORIGEN):
        raise SystemExit(f'No existe el origen: {ORIGEN}')
    if os.path.isdir(DESTINO):
        shutil.rmtree(DESTINO)
    os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
    shutil.copytree(ORIGEN, DESTINO)
    log(f'  copia de trabajo regenerada: {DESTINO}')


def digest(path):
    """Huella comparable de un .xlsx (valores, formatos, relleno, merges, DV)."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked))
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted(f'{dv.type}:{dv.formula1}:{dv.sqref}'
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted(str(r.sqref) for r in ws.conditional_formatting),
            'area': str(ws.print_area),
            'titulos': str(ws.print_title_rows),
        }
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append(f'{fichero}:{hoja}: hoja sólo en una pasada')
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'area', 'titulos'):
            if ha[k] != hb[k]:
                fuera.append(f'{fichero}:{hoja}: cambia {k} '
                             f'({ha[k]} → {hb[k]})'[:400])
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append(f'{fichero}:{hoja}!{coord}: '
                             f'{ca.get(coord)} → {cb.get(coord)}')
    return fuera


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, grupos, informe_global):
    path = os.path.join(carpeta, fname)
    wb, creado = abrir_o_crear(carpeta, fname, grupos)
    cambios = []
    motor.REGISTRO = []
    if creado:
        cambios.insert(0, 'fichero CREADO desde cero por el grupo (no existe '
                          f'en {ORIGEN})')

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []):
            g.pre(wb, fname, cambios)

    motor.aplicar(wb, fname, cambios)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []):
            g.post(wb, fname, cambios, [])

    motor.cerrar(wb, fname, cambios)
    restos = motor.restos_prohibidos(wb)
    wb.save(path)

    registro = list(motor.REGISTRO)
    informe_global.append({'fichero': fname, 'creado': creado,
                           'cambios': cambios,
                           'formulas_nuevas': len(registro),
                           'restos_normativa_derogada': restos})
    return registro, restos


def abrir_o_crear(carpeta, fname, grupos):
    """Libro existente, o libro NUEVO si un grupo lo declara en `NUEVOS`.

    El grupo C aporta cuatro registros (16-19) que NO existen en
    `astro-site/public/dl/pack-appcc`: el análisis de peligros los citaba en su
    columna «Registro asociado» y no había ningún fichero detrás. Se crean aquí,
    dentro de la carpeta de TRABAJO, y el informe los marca con `creado: true`.

    La 2.ª pasada (idempotencia) ya los encuentra en disco y los carga como
    cualquier otro, que es justo lo que hace que la prueba valga: el grupo tiene
    que llegar al mismo resultado partiendo de un libro vacío y partiendo del
    fichero que él mismo escribió.
    """
    path = os.path.join(carpeta, fname)
    if os.path.isfile(path):
        return openpyxl.load_workbook(path), False
    for g in grupos:
        if fname in getattr(g, 'NUEVOS', []):
            return g.crear(fname), True
    raise SystemExit(
        f'No existe {path} y ningún grupo cargado lo declara en NUEVOS. '
        'Si es un registro nuevo, añádelo a `NUEVOS` y a `crear()` de su grupo.')


def ficheros_a_tocar(grupos):
    nombres = list(motor.FICHEROS)
    for g in grupos:
        for f in getattr(g, 'FICHEROS', []):
            if f not in nombres:
                nombres.append(f)
    return nombres


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        fuera.append((n, r.stdout.strip(), r.returncode))
        log('    ' + r.stdout.strip())
        if r.returncode != 0:
            log('    ERROR inject_cache: ' + r.stderr[-400:])
    return fuera


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA debe tener valor cacheado, salvo las que devuelven ""
    por diseño (su texto contiene una cadena vacía)."""
    fallos, vacias, ok = [], 0, 0
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        wbv = openpyxl.load_workbook(path, data_only=True)
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append(f'{fname}:{hoja}!{coord}: hoja ausente')
                continue
            v = wbv[hoja][coord].value
            if v is None:
                if '""' in formula:
                    vacias += 1
                else:
                    fallos.append(f'{fname}:{hoja}!{coord}: sin cache '
                                  f'({formula[:70]})')
            else:
                ok += 1
    return {'con_valor': ok, 'vacias_por_diseno': vacias, 'fallos': fallos}


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}: {e}'


def _json_safe(v):
    """Las entradas de un caso pueden ser fechas u horas (los registros 17, 18
    y BONUS-01 restan fechas), y `json.dump` no las serializa. Se guardan en
    ISO, que es lo que hay que poder leer en el informe."""
    if isinstance(v, (datetime.date, datetime.time, datetime.datetime)):
        return v.isoformat()
    return v


def fuera_de_limite(carpeta, grupos):
    """§6 — un dato FUERA de límite tiene que cambiar el veredicto.

    Se hace sobre COPIAS desechables: la prueba no puede dejar temperaturas de
    mentira dentro del entregable. Cada grupo declara sus casos en
    `CASOS_LIMITE`, así que añadir un registro nuevo (16-19 del grupo C) es
    añadir una entrada a esa lista, no tocar este orquestador.
    """
    from pycel import ExcelCompiler
    os.makedirs(DEMOS, exist_ok=True)
    fuera, fallos = [], []
    casos = [c for g in grupos for c in getattr(g, 'CASOS_LIMITE', [])]
    for i, c in enumerate(casos):
        path = os.path.join(carpeta, c['fichero'])
        if not os.path.isfile(path):
            continue
        dst = os.path.join(DEMOS, f"limite-{i:02d}-{c['fichero']}")
        shutil.copy2(path, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb[c['hoja']]
        base_ref = f"'{c['hoja']}'!{c['salida']}"
        antes = openpyxl.load_workbook(path, data_only=True)[
            c['hoja']][c['salida']].value
        for coord, valor in c['entradas'].items():
            ws[coord] = valor
        wb.save(dst)
        obtenido = _ev(ExcelCompiler(filename=dst), base_ref)
        ok = obtenido == c['esperado']
        fuera.append({
            'ref': f"{c['fichero']}:{c['hoja']}:{c['salida']}",
            'lectura': c['lectura'],
            'entradas': {f"{c['hoja']}!{k}": _json_safe(v)
                         for k, v in c['entradas'].items()},
            'veredicto_con_la_fila_vacia': _json_safe(antes),
            'esperado': c['esperado'], 'obtenido': obtenido, 'ok': ok,
            'copia_desechable': dst})
        if not ok:
            fallos.append(f"fuera de límite {c['fichero']}:{c['hoja']}!"
                          f"{c['salida']} ({c['lectura']}): esperaba "
                          f"{c['esperado']!r}, dio {obtenido!r}")
    return fuera, fallos


# ==========================================================================
# TEC-09 — gate sobre el XML CRUDO
# ==========================================================================
# Toda la cadena de verificación escribe con openpyxl y comprueba releyendo con
# openpyxl (o evaluando con pycel): un round-trip de la misma librería contra sí
# misma, estructuralmente CIEGO a lo que Excel repara o descarta. Así se coló el
# «=» inicial en las 195 reglas del semáforo (TEC-01) con el pipeline en verde.
# Estos tres asserts miran el zip por dentro.
RX_CFRULE = re.compile(rb'<cfRule[^>]*>.*?</cfRule>', re.S)
RX_FORMULA = re.compile(rb'<formula>(.*?)</formula>', re.S)
RX_MERGE = re.compile(rb'<mergeCell ref="([^"]+)"')
RX_VACIO = re.compile(rb'(<c\b[^>]*>)(<f[^>]*>.*?</f>)(?:<v></v>|<v\s*/>)(</c>)', re.S)


def limpiar_xml(carpeta, nombres):
    """Quita el `<v></v>` vacío que openpyxl deja en las celdas de fórmula que
    devuelven cadena vacía. Un `<c>` con `<f>` y sin `<v>` es legal (Excel
    recalcula al abrir: openpyxl escribe `fullCalcOnLoad`), mientras que un
    `<v>` vacío SIN `t="str"` declara un valor numérico que no existe.

    Se hace aquí y no en `inject_cache.py`, que es utilidad compartida por los
    44 productos: el radio de acción de un cambio ahí es todo el catálogo.
    """
    tocadas = 0
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        with zipfile.ZipFile(path) as z:
            partes = [(i, z.read(i.filename)) for i in z.infolist()]
        nuevas, cambio = [], False
        for info, data in partes:
            if info.filename.startswith('xl/worksheets/') and RX_VACIO.search(data):
                data, n_sub = RX_VACIO.subn(rb'\1\2\3', data)
                tocadas += n_sub
                cambio = True
            nuevas.append((info, data))
        if not cambio:
            continue
        tmp = path + '.tmp'
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for info, data in nuevas:
                zout.writestr(info, data)
        os.replace(tmp, path)
    return tocadas


def gate_xml(carpeta, nombres):
    fallos, revisados = [], 0
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        with zipfile.ZipFile(path) as z:
            for nombre in z.namelist():
                if not nombre.startswith('xl/worksheets/'):
                    continue
                data = z.read(nombre)
                for regla in RX_CFRULE.findall(data):
                    for f in RX_FORMULA.findall(regla):
                        revisados += 1
                        if f.lstrip().startswith(b'='):
                            fallos.append(
                                f'{n}:{nombre}: cfRule con «=» inicial '
                                f'→ {f[:70].decode("utf-8", "replace")}')
                for ref in RX_MERGE.findall(data):
                    if b':' not in ref:
                        fallos.append(f'{n}:{nombre}: mergeCell de una sola '
                                      f'celda ({ref.decode()})')
                for m in RX_VACIO.finditer(data):
                    fallos.append(f'{n}:{nombre}: celda de fórmula con '
                                  '<v></v> vacío y sin t="str"')
    return {'cfrule_formulas_revisadas': revisados,
            'fallos': fallos[:40], 'n_fallos': len(fallos)}


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-1500:])
    salida = (r.stdout.strip() or r.stderr.strip()).splitlines()
    return {'exit': r.returncode, 'salida': salida[-8:]}


def inventario(carpeta, nombres):
    """Foto de lo que queda: hojas, filas de registro, DV, CF y fórmulas."""
    fuera = []
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        hojas = []
        for ws in wb.worksheets:
            hojas.append({
                'hoja': ws.title,
                'dim': ws.dimensions,
                'dv': len(ws.data_validations.dataValidation),
                'dv_con_error': sum(
                    1 for dv in ws.data_validations.dataValidation
                    if dv.showErrorMessage),
                'cf': len(list(ws.conditional_formatting)),
                'print_area': ws.print_area,
                'title_rows': ws.print_title_rows,
                'a4': ws.page_setup.paperSize == 9,
                'formulas': sum(1 for row in ws.iter_rows() for c in row
                                if isinstance(c.value, str)
                                and c.value.startswith('=')),
            })
        fuera.append({'fichero': n, 'hojas': hojas,
                      'subject': wb.properties.subject,
                      'creator': wb.properties.creator})
    return fuera


# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 del Pack de Plantillas APPCC')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    ap.add_argument('--solo', default='a',
                    help='grupos a aplicar: a, b, c o a,b')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--sin-idempotencia', action='store_true')
    args = ap.parse_args()

    if not args.dry_run and os.environ.get('PACK_APPCC_APPLY') != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en '
            f'{ORIGEN}, que la SPEC declara intocable hasta que la ronda 2 dé '
            'verde. Usa --dry-run (o PACK_APPCC_APPLY=1 si eres el '
            'orquestador y ya tienes el visto bueno).')

    carpeta = DESTINO if args.dry_run else ORIGEN
    respaldo = None
    if args.dry_run:
        preparar_copia()
    else:
        respaldo = os.path.join(
            SCRATCH, 'pack-appcc.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(ORIGEN, respaldo)
        log(f'  respaldo previo de los entregables: {respaldo}')

    grupos = []
    pedidos = [g.strip().lower() for g in args.solo.split(',') if g.strip()]
    for letra in pedidos:
        try:
            grupos.append(importlib.import_module(f'grupo_{letra}'))
            log(f'  grupo_{letra} cargado')
        except ImportError:
            log(f'  grupo_{letra} no existe todavía — se omite')

    nombres = ficheros_a_tocar(grupos)
    informe_ficheros, registros, restos_todos = [], {}, []
    log(f'\n== 1/7 · post-proceso de {len(nombres)} ficheros ==')
    for fname in nombres:
        nuevo = not os.path.isfile(os.path.join(carpeta, fname))
        registros[fname], restos = procesar(carpeta, fname, grupos,
                                            informe_ficheros)
        restos_todos += [f'{fname}: {r}' for r in restos]
        log(f'  {fname}: {len(registros[fname])} fórmulas'
            + ('   ← CREADO desde cero' if nuevo else ''))

    # ---- idempotencia ---------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 2/7 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(IDEM):
            shutil.rmtree(IDEM)
        shutil.copytree(carpeta, IDEM)
        antes = {n: digest(os.path.join(carpeta, n)) for n in nombres}
        for fname in nombres:
            procesar(IDEM, fname, grupos, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(IDEM, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:40]}
        log(f'  diferencias 1.ª vs 2.ª pasada: {len(difs)}')
        for d in difs[:10]:
            log('    ' + d)

    # ---- cache ----------------------------------------------------------
    log('\n== 3/7 · inject_cache (al final del todo) ==')
    cache = inject_cache(carpeta, nombres)

    log('\n== 4/7 · verificación data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log(f"  con valor: {ver['con_valor']} · \"\" por diseño: "
        f"{ver['vacias_por_diseno']} · fallos: {len(ver['fallos'])}")
    for f in ver['fallos'][:10]:
        log('    ' + f)

    log('\n== 4b/8 · limpieza y gate del XML crudo (TEC-09) ==')
    vacias = limpiar_xml(carpeta, nombres)
    gxml = gate_xml(carpeta, nombres)
    log(f"  <v></v> vacíos retirados: {vacias} · fórmulas de cfRule "
        f"revisadas: {gxml['cfrule_formulas_revisadas']} · fallos: "
        f"{gxml['n_fallos']}")
    for f in gxml['fallos'][:10]:
        log('    ' + f)

    log('\n== 5/7 · pycel con inputs FUERA de límite (§6) ==')
    lim, lim_fallos = fuera_de_limite(carpeta, grupos)
    for c in lim:
        log(f"  {c['ref']}: {c['lectura']} → {c['obtenido']!r} "
            f"{'OK' if c['ok'] else 'FALLA (esperaba ' + repr(c['esperado']) + ')'}")

    log('\n== 6/7 · censo-entregables --fail ==')
    cen = censo(carpeta)

    log('\n== 7/7 · demostraciones de los grupos (§6) ==')
    demos, fallos = {}, []
    for g in grupos:
        if hasattr(g, 'demos'):
            propias = g.demos(carpeta, ORIGEN, DEMOS)
            fallos += propias.pop('fallos', [])
            demos[g.__name__] = propias
            for c in propias.get('casos_spec_6', []):
                log(f"  {c['ref']}: {c['entradas']} → {c['obtenido']!r} "
                    f"{'OK' if c['ok'] else 'FALLA'}")
            log(f"  {g.__name__}: normas derogadas vivas = "
                f"{propias.get('normativa_ocurrencias_totales')}")

    if idem.get('diferencias'):
        fallos.append(f"idempotencia: {idem['diferencias']} diferencias")
    if ver['fallos']:
        fallos.append(f"cache: {len(ver['fallos'])} fórmulas sin valor")
    fallos += lim_fallos
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió ' + str(cen['exit']))
    if gxml['n_fallos']:
        fallos.append(f"XML crudo: {gxml['n_fallos']} defectos que openpyxl "
                      'acepta y Excel repara (TEC-09)')
    fallos += [f'normativa derogada viva: {r}' for r in restos_todos]

    informe = {
        'producto': 'pack-appcc',
        'version': '2.0',
        'spec': SPEC,
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'carpeta_origen': ORIGEN,
        'carpeta_trabajo': carpeta,
        'grupos': pedidos,
        'ficheros_creados': [f['fichero'] for f in informe_ficheros
                             if f.get('creado')],
        'ficheros': informe_ficheros,
        'inventario': inventario(carpeta, nombres),
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'data_only_formulas_nuevas': ver,
            'xml_crudo': dict(gxml, v_vacios_retirados=vacias),
            'pycel_fuera_de_limite': lim,
            'censo_entregables': cen,
        },
        'demostraciones': demos,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1)
        log(f'\ninforme → {args.json}')

    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'TODO VERDE (idempotencia, cache, pycel, censo, §6)'))
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (BÓRRALO sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
