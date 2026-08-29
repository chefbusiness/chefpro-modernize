#!/usr/bin/env python3
"""
grupo_c.py — §4 de `pack-appcc-v2-SPEC.md`: alérgenos, los CUATRO registros
nuevos y los dos bonos.

  · 08 matriz de alérgenos — los 8 platos de ejemplo con su declaración REAL en
    las 14 columnas (DOM-02/TEC-03/COM-17, altas), columna «Verificado» con
    semáforo, «Especificación» del cereal y del fruto de cáscara concretos
    (DOM-20), cabeceras con ajuste de texto (TEC-18) y rangos a 200 filas
    (DOM-19).
  · 16 cocción / regeneración — NUEVO. La columna J del análisis de peligros
    apuntaba a un registro que no existía (DOM-01/TEC-04/COM-04, altas).
  · 17 enfriamiento / descongelación — NUEVO, dos pestañas. Ídem.
  · 18 congelación preventiva de anisakis — NUEVO. Obligación con sanción en
    España (RD 1021/2022, art. 8.1, que derogó el RD 1420/2006, y Reg. (CE)
    853/2004 Anexo III, Secc. VIII, Cap. III.D) que el pack no cubría por
    ninguna parte (DOM-06, alta).
  · 19 verificación de termómetros — NUEVO. Todo el pack decide con sondas que
    nadie contrastaba (DOM-10/COM-07).
  · BONUS-01 formación — «Válido hasta», «Firma del empleado» y Estado con
    TODAY() (VIGENTE / RENOVAR / CADUCADO) + contadores (DOM-33/TEC-28/COM-25).
  · BONUS-02 protocolo — PASO 3 reescrito: la obligación legal es notificar a
    la autoridad sanitaria de la comunidad autónoma (art. 19 Reg. 178/2002),
    no llamar al 112 (DOM-21). A4 con nota de ampliación a A3 (DOM-30/TEC-21).

Los cuatro registros nuevos NO existen en `astro-site/public/dl/pack-appcc`:
este módulo los CREA (`NUEVOS` + `crear()`), `main.py` los guarda en la carpeta
de trabajo y los registra en el informe. Todo lo demás sigue el mismo contrato
que `grupo_a.py`: `pre()` para lo estructural, `post()` para la rejilla entera,
`demos()` y `CASOS_LIMITE` para §6.

DECISIONES TÉCNICAS que conviene no revertir sin leer el motivo:

 1. **COUNTBLANK está PROHIBIDO aquí.** La SPEC propone
    `=IF(COUNTBLANK(D6:Q6)>0,"⚠ SIN VERIFICAR","Completo")`, pero pycel NO
    implementa COUNTBLANK (medido: `UnknownFunction … COUNTBLANK is in the
    "Statistical" group`). La celda se quedaría sin cache, el visor móvil la
    enseñaría en blanco y —peor— el gate del §6 no podría demostrar que la
    matriz avisa. El equivalente exacto, porque la validación sólo admite S/T/N,
    es contar las tres marcas y compararlas con 14.
 2. **Las fechas se guardan como FECHA, nunca como texto.** El Estado de
    BONUS-01 compara con TODAY(): en Excel cualquier texto es mayor que
    cualquier número, así que un «01/03/2020» tecleado como texto devolvería
    VIGENTE para un certificado caducado hace años. Por eso las columnas de
    fecha llevan `dv_fecha()` (Excel rechaza el texto) y las fórmulas se guardan
    además con `ISNUMBER`, que es el cinturón por si alguien pega valores.
 3. **TODAY() se cachea con la fecha de generación.** Es aceptable —Excel
    recalcula al abrir— y queda anotado en el informe; lo que NO sería
    aceptable es que la fórmula no existiera.
 4. Los ocho platos de ejemplo llevan «(ejemplo)» EN EL NOMBRE, no en una
    columna de observaciones (la matriz no tiene). Una carta de alérgenos es lo
    único del pack que se le enseña a un comensal alérgico: tiene que ser
    imposible confundir el ejemplo con la carta real aunque se imprima tal cual.
"""
import contextlib
import datetime
import os
import shutil

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import motor

F_08 = '08-matriz-alergenos.xlsx'
F_16 = '16-registro-coccion-regeneracion.xlsx'
F_17 = '17-registro-enfriamiento-descongelacion.xlsx'
F_18 = '18-registro-congelacion-anisakis.xlsx'
F_19 = '19-verificacion-termometros.xlsx'
F_B1 = 'BONUS-01-registro-formacion.xlsx'
F_B2 = 'BONUS-02-protocolo-alerta-alimentaria.xlsx'

FICHEROS = [F_08, F_16, F_17, F_18, F_19, F_B1, F_B2]

#: Ficheros que este grupo CREA desde cero (no existen en `dl/pack-appcc`).
NUEVOS = [F_16, F_17, F_18, F_19]

#: Pestañas de datos de cada fichero nuevo, en orden. `crear()` añade siempre
#: «Instrucciones» delante, como el resto del pack.
HOJAS_NUEVAS = {
    F_16: ['Cocción y Regeneración'],
    F_17: ['Enfriamiento', 'Descongelación'],
    F_18: ['Congelación Anisakis'],
    F_19: ['Verificación Termómetros'],
}

#: Metadata de los ficheros nuevos, con el mismo patrón que los 17 heredados
#: («<Título legible> · Pack de Plantillas APPCC»). El `subject` lo pone
#: `motor.set_metadata`.
TITULOS = {
    F_16: 'Registro de Cocción y Regeneración',
    F_17: 'Registro de Enfriamiento y Descongelación',
    F_18: 'Registro de Congelación Preventiva de Anisakis',
    F_19: 'Verificación de Termómetros y Sondas',
}

#: Formatos de fecha/hora. Van APARTE de los del motor porque aquí no son
#: cosmética: tres veredictos (descongelación, anisakis y formación) restan
#: fechas, y una celda con formato de TEXTO ('@') hace que Excel guarde como
#: texto lo que el usuario teclea. Con eso, la resta devuelve #VALUE! y la
#: comparación con TODAY() sale al revés (en Excel cualquier texto es mayor que
#: cualquier número). Por eso toda columna que alimente una fórmula lleva su
#: formato numérico de fecha, y no el '@' por defecto de `_pintar_fila`.
FMT_FECHA = 'DD/MM/YYYY'
FMT_HORA = 'HH:MM'
FMT_FECHAHORA = 'DD/MM/YYYY HH:MM'

KEYWORDS = 'pack appcc, AI Chef Pro'
CATEGORIA = 'AI Chef Pro · Productos digitales'
DESCRIPCION = 'aichef.pro/pack-appcc'


# ==========================================================================
# Utilidades locales
# ==========================================================================
def dv_fecha(ws, rangos, titulo, error, desde='2000-01-01', hasta='2100-12-31'):
    """Validación de FECHA que rechaza el texto (§ decisión técnica 2).

    `motor` sólo trae `dv_lista` y `dv_decimal`; aquí hace falta la tercera,
    porque tres de los cinco veredictos nuevos (formación, descongelación,
    anisakis) se calculan restando fechas y una fecha escrita como texto no
    resta: devuelve #VALUE! o, peor, se compara como texto y sale conforme.
    """
    dv = DataValidation(
        type='date', operator='between',
        formula1=f'DATE({desde[:4]},{int(desde[5:7])},{int(desde[8:10])})',
        formula2=f'DATE({hasta[:4]},{int(hasta[5:7])},{int(hasta[8:10])})',
        allow_blank=True, showErrorMessage=True, errorStyle='stop',
        errorTitle=titulo, error=error)
    ws.add_data_validation(dv)
    for r in rangos:
        dv.add(r)
    return dv


def dv_hora(ws, rangos, titulo, error):
    """Validación de HORA (TEC-05, ronda 2).

    Las únicas columnas de las que depende el veredicto de los PCC nuevos vía
    `ISNUMBER()` eran precisamente las únicas SIN validación: la fecha (A) sí la
    tenía, las horas y las fechas-y-hora no. Si el usuario escribe «16.00»,
    «16h» o pega desde otra hoja y la celda queda como TEXTO, `ISNUMBER` falla,
    la columna de tiempo devuelve "" y el Estado se queda en blanco — sin aviso,
    sin color y sin que ningún gate lo note. El autor conocía el riesgo y lo
    avisaba en prosa en las Instrucciones, que es justo lo que la §1.2 dice que
    hay que dejar de hacer con una nota y empezar a hacer con una DV.
    """
    dv = DataValidation(
        type='time', operator='between',
        formula1='TIME(0,0,0)', formula2='TIME(23,59,59)',
        allow_blank=True, showErrorMessage=True, errorStyle='stop',
        errorTitle=titulo, error=error)
    ws.add_data_validation(dv)
    for r in rangos:
        dv.add(r)
    return dv


def dv_fechahora(ws, rangos, titulo, error):
    """Validación de FECHA Y HORA (TEC-05). Es una DV de fecha: Excel compara el
    número de serie, y un «05/09/2026 09:00» cae dentro del rango igual que la
    fecha sola. Lo que rechaza es el TEXTO, que es lo que rompe la resta."""
    return dv_fecha(ws, rangos, titulo, error)


def _titulo(ws, texto, ncols):
    ws['A1'] = texto
    ws['A1'].font = Font(bold=True, size=16)
    ref = f'A1:{openpyxl.utils.get_column_letter(ncols)}1'
    if ref not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells(ref)


def _pintar_fila(ws, fila, ncols, calc=(), texto_izq=(), num=None, fmt=None):
    """Fila plantilla: verde donde se escribe, gris-calculada donde no."""
    for col in range(1, ncols + 1):
        cel = ws.cell(row=fila, column=col)
        cel.value = None
        if col in calc:
            motor.calculada(cel, (fmt or {}).get(col))
        elif num and col in num:
            motor.verde(cel, (fmt or {}).get(col, motor.FMT_TEMP))
        elif col in texto_izq:
            motor.verde(cel, align='left')
        else:
            motor.verde(cel, (fmt or {}).get(col, '@'))


def _pie(ws, fila, ncols, extra=()):
    """Notas de cierre comunes: lo específico, la conservación y la marca."""
    for i, texto in enumerate(extra):
        motor.nota(ws, fila + i, texto, ncols=ncols)
    motor.nota(ws, fila + len(extra), motor.CONSERVACION, ncols=ncols)
    motor.nota(ws, fila + len(extra) + 1, motor.MARCA, ncols=ncols)


# ==========================================================================
# 08 — Matriz de alérgenos
# ==========================================================================
# DOM-20: el Anexo II del Reg. (UE) 1169/2011 no obliga a declarar «gluten»,
# obliga a declarar el CEREAL (trigo, centeno, cebada, avena, espelta, kamut) y
# el FRUTO DE CÁSCARA concreto (almendra, avellana, nuez, anacardo…). Por eso
# esas dos cabeceras piden la especie y hay una columna donde escribirla.
ALERGENOS_08 = [
    'Cereales con gluten (indicar cuál)', 'Crustáceos', 'Huevos', 'Pescado',
    'Cacahuetes', 'Soja', 'Lácteos', 'Frutos de cáscara (indicar cuál)',
    'Apio', 'Mostaza', 'Sésamo', 'Sulfitos (SO2)', 'Altramuces', 'Moluscos',
]
CAB_08 = ['Nº', 'Plato / Producto', 'Categoría'] + ALERGENOS_08 + [
    'Especificación (cereal o fruto de cáscara concreto)', 'Verificado']
ANCHOS_08 = [6, 40, 15] + [12] * 14 + [42, 18]

COL_D, COL_Q, COL_R, COL_S = 4, 17, 18, 19        # D..Q alérgenos, R, S
F0_08, F1_08 = 6, 205                             # 200 filas (DOM-19)
TOT_08, SINVER_08 = 207, 208
CATS_08 = ('Entrantes,Primeros,Segundos,Postres,Bebidas,Tapas,Desayunos,'
           'Infantil,Otros')

# DOM-02/TEC-03/COM-17 (altas): en v1.1 las 112 casillas de los 8 platos de
# ejemplo estaban VACÍAS y la propia leyenda define el vacío como «no
# verificado». Un local que imprimiera la matriz sin borrar las filas demo
# entregaba a un cliente celíaco unas croquetas y una tarta de queso con cero
# alérgenos declarados, en el documento que el propio fichero llama OBLIGATORIO.
# Orden de las marcas = orden de ALERGENOS_08 (14 caracteres, S/T/N).
PLATOS_08 = [
    ('Ensalada César', 'Entrantes', 'SNSSNTSNNSNNNN',
     'Cereal: trigo (picatostes de pan). La soja va en trazas por la salsa '
     'Worcestershire.'),
    ('Croquetas de jamón', 'Entrantes', 'SNTNNNSNNNNNNN',
     'Cereal: trigo (harina de la bechamel y pan rallado del rebozado).'),
    ('Sopa de pescado', 'Primeros', 'TSNSNNNNSNNSNT',
     'Cereal: trigo, en trazas (picatostes de acompañamiento).'),
    ('Paella mixta', 'Primeros', 'TSNSNNNNNNNTNS',
     'Cereal: trigo, en trazas (se elabora pasta en la misma cocina).'),
    ('Solomillo con salsa Pedro Ximénez', 'Segundos', 'NNNNNNSNNNNSNN', None),
    ('Merluza a la vasca', 'Segundos', 'SNSSNNNNNNNSNS',
     'Cereal: trigo (harina para enharinar la merluza).'),
    ('Tarta de queso', 'Postres', 'SNSNNNSTNNNNNN',
     'Cereal: trigo (base de galleta). Fruto de cáscara: almendra, en trazas '
     '(obrador compartido).'),
    ('Helado de vainilla', 'Postres', 'NNSNTTSTNNNNNN',
     'Fruto de cáscara: avellana y almendra, en trazas (misma máquina '
     'heladora).'),
]

SIN_VERIFICAR = '⚠ SIN VERIFICAR'
FALTA_ESPECIE = '⚠ FALTA ESPECIE'


def _formula_verificado(fila):
    """Tres COUNTIF en vez de COUNTBLANK (decisión técnica 1).

    La validación de D:Q sólo admite S, T o N y RECHAZA lo demás (§1.2), así
    que «cuántas de las 14 llevan una de las tres marcas» y «cuántas están
    rellenas» son el mismo número. Comparado con 14, dice si la fila está
    declarada entera.

    El tercer nivel es DOM-20: una fila con gluten o frutos de cáscara marcados
    y la especie sin escribir NO está declarada conforme al Anexo II, aunque
    tenga las 14 casillas puestas. No es rojo (la información está a medias, no
    ausente): es ámbar.
    """
    marcas = '+'.join(f'COUNTIF($D{fila}:$Q{fila},"{m}")' for m in 'STN')
    especie = (f'AND(OR($D{fila}="S",$D{fila}="T",$K{fila}="S",'
               f'$K{fila}="T"),$R{fila}="")')
    return (f'=IF($B{fila}="","",'
            f'IF({marcas}<14,"{SIN_VERIFICAR}",'
            f'IF({especie},"{FALTA_ESPECIE}","Completo")))')


def _post_08(wb, fname, cambios):
    ws = motor.hoja(wb, 'Matriz Alérgenos')

    _titulo(ws, 'Matriz de Alérgenos — Carta Completa', COL_S)
    ws['A2'] = ('Establecimiento: ________________________________    '
                'Fecha de actualización: ___/___/______')
    ws['A3'] = ('Leyenda:  S = Contiene  |  T = Trazas  |  N = No contiene  |  '
                '(vacío) = No verificado — una casilla vacía NO significa que '
                'el plato no lleve ese alérgeno.')
    ws['A3'].font = Font(size=10, italic=True)
    ws['A4'] = ('⚠ Las 8 primeras filas son EJEMPLOS y llevan «(ejemplo)» en el '
                'nombre: bórralas y escribe tu carta real antes de imprimir '
                'esta matriz o enseñársela a un cliente.')
    ws['A4'].font = Font(size=10, bold=True, color=motor.SEM_ROJ_FG)
    ws['A4'].fill = PatternFill('solid', fgColor=motor.SEM_ROJ_BG)
    for r in (3, 4):
        ref = f'A{r}:{openpyxl.utils.get_column_letter(COL_S)}{r}'
        if ref not in [str(m) for m in ws.merged_cells.ranges]:
            ws.merge_cells(ref)

    motor.cabecera(ws, 5, CAB_08, ANCHOS_08)
    # TEC-18: «Frutos de cáscara» se leía «Frutos cásca» en pantalla y en papel
    # (14 caracteres en una columna de ancho 12, sin ajuste de texto y con la
    # vecina ocupada). `motor.cabecera` ya activa el ajuste; lo que faltaba es
    # el alto para que las cuatro líneas quepan.
    ws.row_dimensions[5].height = 64

    _pintar_fila(ws, F0_08, COL_S, calc=(1, COL_S), texto_izq=(2, COL_R))
    ws.cell(row=F0_08, column=1).value = None
    f_num = f'=IF($B{F0_08}="","",ROW()-{F0_08 - 1})'
    f_ver = _formula_verificado(F0_08)
    ws.cell(row=F0_08, column=1).value = f_num
    ws.cell(row=F0_08, column=COL_S).value = f_ver
    motor.reg(ws, f'A{F0_08}', f_num)
    motor.reg(ws, f'S{F0_08}', f_ver)
    motor.replicar_filas(ws, F0_08, F0_08, F1_08, ncols=COL_S, alto=20)

    # §1.2 — DV que RECHAZA. En v1.1 las dos listas tenían
    # showErrorMessage=False: se podía teclear «si», «x» o «??» y el COUNTIF de
    # totales —que compara por igualdad exacta— no lo contaba, sin avisar.
    motor.dv_lista(ws, '"S,T,N"', [f'D{F0_08}:Q{F1_08}'], 'Alérgeno',
                   'Marca S (contiene), T (trazas) o N (no contiene). '
                   'Cualquier otro texto no lo cuenta el resumen ni lo '
                   'reconoce la columna «Verificado».')
    motor.dv_lista(ws, f'"{CATS_08}"', [f'C{F0_08}:C{F1_08}'], 'Categoría',
                   'Elige una categoría de la lista.')

    motor.semaforo(ws, f'S{F0_08}:S{F1_08}',
                   extra_ambar=(FALTA_ESPECIE,), extra_rojo=(SIN_VERIFICAR,))

    for i, (nombre, cat, marcas, especie) in enumerate(PLATOS_08):
        fila = F0_08 + i
        ws.cell(row=fila, column=2).value = f'{nombre} (ejemplo)'
        ws.cell(row=fila, column=3).value = cat
        for j, marca in enumerate(marcas):
            ws.cell(row=fila, column=COL_D + j).value = marca
        if especie:
            ws.cell(row=fila, column=COL_R).value = especie

    # DOM-19: los COUNTIF de v1.1 estaban clavados en la fila 45 y la landing
    # promete «todos los platos de tu carta». A partir del plato 41 la matriz
    # dejaba de contar sin un solo aviso.
    ws.cell(row=TOT_08, column=2,
            value='TOTAL de platos que contienen el alérgeno (S) o pueden '
                  'contenerlo (T):').font = Font(bold=True, size=11)
    for col in range(COL_D, COL_Q + 1):
        letra = openpyxl.utils.get_column_letter(col)
        f = (f'=COUNTIF({letra}{F0_08}:{letra}{F1_08},"S")'
             f'+COUNTIF({letra}{F0_08}:{letra}{F1_08},"T")')
        ws.cell(row=TOT_08, column=col).value = f
        motor.calculada(ws.cell(row=TOT_08, column=col), motor.FMT_ENT)
        ws.cell(row=TOT_08, column=col).font = Font(bold=True, size=11)
        motor.reg(ws, f'{letra}{TOT_08}', f)

    ws.cell(row=SINVER_08, column=2,
            value='Platos SIN la declaración completa (no se pueden publicar):'
            ).font = Font(bold=True, size=11)
    f = (f'=COUNTIF($S${F0_08}:$S${F1_08},"{SIN_VERIFICAR}")'
         f'+COUNTIF($S${F0_08}:$S${F1_08},"{FALTA_ESPECIE}")')
    ws.cell(row=SINVER_08, column=COL_D).value = f
    motor.calculada(ws.cell(row=SINVER_08, column=COL_D), motor.FMT_ENT)
    ws.cell(row=SINVER_08, column=COL_D).font = Font(bold=True, size=12)
    motor.reg(ws, f'D{SINVER_08}', f)
    motor.cf_formula(ws, f'D{SINVER_08}', f'=D{SINVER_08}>0')

    _pie(ws, 210, COL_S, extra=[
        'La columna «Verificado» avisa sola: «⚠ SIN VERIFICAR» mientras falte '
        'alguna de las 14 casillas y «⚠ FALTA ESPECIE» cuando hay gluten o '
        'frutos de cáscara marcados y no se ha escrito de cuál se trata. Sólo '
        'las filas que dicen «Completo» se pueden publicar.',
        'Esta carta de alérgenos debe estar disponible para cualquier cliente '
        'que la solicite (Reg. (UE) 1169/2011 y RD 126/2015). Actualízala cada '
        'vez que cambie la carta, la receta o el proveedor de un ingrediente.',
        'La matriz admite 200 platos: los desplegables, la columna '
        '«Verificado» y los totales llegan hasta la fila 205.',
    ])
    motor.IMPRESION[(fname, ws.title)] = (5, True, 'D6')  # TEC-08

    motor.escribir_instrucciones(wb, 'Matriz de Alérgenos — Carta Completa', [
        ('h', 'Cómo usar esta plantilla'),
        ('b', 'Escribe TODOS los platos de tu carta en la columna «Plato / '
              'Producto» y elige su categoría. La numeración se pone sola.'),
        ('b', 'Para cada plato, marca las 14 columnas de alérgeno con S '
              '(contiene), T (trazas) o N (no contiene). Las tres son la única '
              'respuesta admitida: no dejes ninguna en blanco.'),
        ('b', 'Una casilla vacía NO quiere decir «no lleva»: quiere decir «no '
              'verificado», y así lo dice la leyenda de la hoja.'),
        ('b', 'La columna «Verificado» se calcula sola y se pinta: «Completo» '
              'en verde, «⚠ FALTA ESPECIE» en ámbar y «⚠ SIN VERIFICAR» en '
              'rojo. Debajo de la tabla hay un contador de platos que todavía '
              'no se pueden publicar.'),
        ('b', 'Las 8 primeras filas son ejemplos reales de alérgeno múltiple y '
              'llevan «(ejemplo)» en el nombre. Bórralas antes de usar la '
              'matriz de verdad.'),
        ('h', 'Por qué hay una columna «Especificación»'),
        ('p', 'El Anexo II del Reglamento (UE) 1169/2011 no obliga a declarar '
              '«gluten» ni «frutos de cáscara» en genérico: obliga a decir el '
              'cereal (trigo, centeno, cebada, avena, espelta, kamut) y el '
              'fruto concreto (almendra, avellana, nuez, anacardo, pistacho, '
              'macadamia, pacana, Brasil). Un celíaco necesita saber si es '
              'avena; un alérgico a la avellana, si la almendra le vale.'),
        ('b', 'Si marcas S o T en «Cereales con gluten» o en «Frutos de '
              'cáscara», escribe la especie en «Especificación». Mientras no lo '
              'hagas, la fila queda en ámbar.'),
        ('h', 'Los 14 alérgenos de declaración obligatoria'),
        ('b', '1. Cereales con gluten | 2. Crustáceos | 3. Huevos | '
              '4. Pescado'),
        ('b', '5. Cacahuetes | 6. Soja | 7. Lácteos | 8. Frutos de cáscara'),
        ('b', '9. Apio | 10. Mostaza | 11. Sésamo | 12. Sulfitos (por encima '
              'de 10 mg/kg, expresados como SO2)'),
        ('b', '13. Altramuces | 14. Moluscos'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Revisa la matriz cada vez que cambies la carta, una receta o el '
              'proveedor de un ingrediente, y al menos una vez al año.'),
        ('b', 'Imprímela en A4 apaisado; la fila de cabecera se repite en cada '
              'página.'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('08: 8 platos con su declaración real en las 14 columnas, '
                   'columnas «Especificación» y «Verificado» con semáforo, '
                   'rangos a 200 filas y cabeceras con ajuste de texto '
                   '(DOM-02/TEC-03/COM-17/DOM-19/DOM-20/TEC-18)')


# ==========================================================================
# 16 — Cocción y regeneración (NUEVO)
# ==========================================================================
CAB_16 = ['Fecha', 'Plato / elaboración', 'Proceso', 'Hora',
          'T.ª en el centro (°C)', 'Tiempo (min)', 'Sonda / termómetro nº',
          'Estado', 'Acción realizada', 'Firma']
ANCHOS_16 = [14, 34, 18, 10, 18, 14, 20, 14, 32, 14]
F0_16, F1_16 = 5, 44


def _post_16(wb, fname, cambios):
    ws = motor.hoja(wb, 'Cocción y Regeneración')
    _titulo(ws, 'Registro de Cocción y Regeneración', 10)
    ws['A2'] = ('Mes: _______________    '
                'Responsable: _________________________________')
    motor.cabecera(ws, 4, CAB_16, ANCHOS_16)

    _pintar_fila(ws, F0_16, 10, calc=(8,), texto_izq=(2, 9),
                 num=(5, 6), fmt={1: FMT_FECHA, 4: FMT_HORA,
                                  5: motor.FMT_TEMP, 6: motor.FMT_ENT})
    # Un solo umbral de temperatura, 75 °C, para cocción Y regeneración: es más
    # estricto que los 65 °C que se venían pidiendo al recalentado y nunca da un
    # OK que no toque.
    #
    # DOM-R2-07 (alta, ronda 2): el límite crítico que declara el plan APPCC
    # para este PCC tiene DOS componentes —«≥75 °C en el centro; en
    # regeneración, alcanzarlos en menos de una hora»— y el registro que lo
    # vigila sólo miraba el primero. La columna «Tiempo (min)» se pedía, se
    # validaba (0-600) y no la leía ninguna fórmula, así que una regeneración
    # que tardaba tres horas en llegar a los 75 °C —la bandeja metida en el
    # horno a media potencia a media mañana— se acreditaba OK en verde. Es el
    # mismo defecto que la ronda 1 cazó en el aceite (DOM-25/TEC-07: un límite
    # declarado que ninguna fórmula mira), corregido allí y reintroducido aquí
    # en el PCC que el propio fichero llama «el más importante de la cocina».
    # `IF($C="Regeneración",…)` sólo se activa cuando hay proceso y tiempo: si
    # falta el tiempo, la fila se decide por la temperatura como antes.
    f = (f'=IF($E{F0_16}="","",'
         f'IF($E{F0_16}<75,"REPETIR",'
         f'IF(AND($C{F0_16}="Regeneración",$F{F0_16}<>"",$F{F0_16}>60),'
         f'"REPETIR","OK")))')
    ws.cell(row=F0_16, column=8).value = f
    motor.reg(ws, f'H{F0_16}', f)
    motor.replicar_filas(ws, F0_16, F0_16, F1_16, ncols=10, alto=20)

    dv_fecha(ws, [motor.rango('A', F0_16, F1_16)], 'Fecha',
             'Escribe una fecha (por ejemplo 05/09/2026).')
    motor.dv_lista(ws, '"Cocción,Regeneración"',
                   [motor.rango('C', F0_16, F1_16)], 'Proceso',
                   'Elige «Cocción» (primera elaboración) o «Regeneración» '
                   '(recalentado de un producto ya cocinado).')
    motor.dv_decimal(ws, [motor.rango('E', F0_16, F1_16)],
                     'Temperatura en el centro',
                     'Introduce la temperatura medida en el CENTRO de la '
                     'pieza, en °C, entre 0 y 250.', minimo=0, maximo=250)
    motor.dv_decimal(ws, [motor.rango('F', F0_16, F1_16)], 'Tiempo',
                     'Introduce el tiempo en minutos, entre 0 y 600.',
                     minimo=0, maximo=600)
    dv_hora(ws, [motor.rango('D', F0_16, F1_16)], 'Hora',
            'Escribe la hora como HORA (por ejemplo 13:45), no como texto.')
    motor.semaforo(ws, motor.rango('H', F0_16, F1_16), extra_rojo=('REPETIR',))

    # OJO: aquí NO se puede sembrar la columna H. Es la del veredicto, y
    # `sembrar` escribe lo que se le pase —incluido un None— encima de la
    # fórmula que acaba de replicar `replicar_filas`. La primera versión traía
    # un `'H': None` y dejaba la fila 5 sin Estado: 39 fórmulas en vez de 40, un
    # fallo que ni el censo ni el gate de cache detectan (una fórmula que no
    # existe no puede estar sin cachear).
    motor.sembrar(ws, 5, {'A': datetime.date(2026, 9, 5),
                          'B': 'Lasaña de carne (bandeja)',
                          'C': 'Cocción', 'D': datetime.time(12, 20),
                          'E': 82, 'F': 45,
                          'G': 'S-01', 'J': 'A.R.'}, marca_col='I')
    motor.sembrar(ws, 6, {'A': datetime.date(2026, 9, 5),
                          'B': 'Croquetas de jamón',
                          'C': 'Regeneración', 'D': datetime.time(13, 45),
                          'E': 78, 'F': 8,
                          'G': 'S-01', 'J': 'M.S.'}, marca_col='I')
    motor.sembrar(ws, 7, {'A': datetime.date(2026, 9, 6),
                          'B': 'Estofado de ternera',
                          'C': 'Regeneración', 'D': datetime.time(13, 10),
                          'E': 70, 'F': 10,
                          'G': 'S-02',
                          # DOM-R2-15: era INC-004, el MISMO número que usa el
                          # 07 para un saco de harina roído. Dos sucesos
                          # distintos bajo un número es lo que impide
                          # reconstruir un caso, y el 11 ya desarrolla el
                          # INC-007 con esta misma regeneración.
                          'I': 'Vuelto a calentar hasta 79 °C · incidencia '
                               'INC-007', 'J': 'M.S.'}, marca_col='I')

    _pie(ws, F1_16 + 2, 10, extra=[
        'Criterio de la hoja: 75 °C o más en el CENTRO de la pieza. Si sale '
        'REPETIR, vuelve a calentar hasta alcanzarlo y anota la acción; nunca '
        'se sirve una elaboración con el veredicto en rojo.',
        'En regeneración el veredicto mira TAMBIÉN el tiempo: si tarda más de '
        '60 minutos en alcanzar los 75 °C, dice REPETIR aunque los alcance. El '
        'producto ha pasado demasiado rato en la zona de peligro.',
        'El RD 1021/2022, que derogó al RD 3484/2000, ya no fija temperaturas '
        'ni tiempos de recalentamiento: los establece y justifica cada '
        'establecimiento en su APPCC (Reg. (CE) 852/2004, art. 5). Los 75 °C '
        'en menos de una hora son el criterio técnico que aplica este pack, y '
        'es el que figura como límite crítico en el análisis de peligros '
        '(fichero 12).',
        'Registro de vigilancia del PCC de cocción del análisis de peligros '
        '(fichero 12). Si hay desviación, abre la incidencia en el registro 11.',
    ])
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    motor.escribir_instrucciones(wb, 'Registro de Cocción y Regeneración', [
        ('h', 'Para qué sirve esta plantilla'),
        ('p', 'Es el registro de vigilancia del punto de control crítico de '
              'cocción y del recalentamiento de comidas preparadas. El análisis '
              'de peligros (fichero 12) apunta aquí: sin esta hoja, el PCC más '
              'importante de la cocina no tenía dónde acreditarse.'),
        ('h', 'Cómo usarla'),
        ('b', 'Una línea por elaboración cocinada o regenerada. No hace falta '
              'medir cada ración: se mide la pieza o la bandeja.'),
        ('b', 'Mide en el CENTRO GEOMÉTRICO de la pieza, con la sonda limpia y '
              'desinfectada, y espera a que la lectura se estabilice.'),
        ('b', 'Anota qué sonda has usado. Su verificación mensual va en el '
              'fichero 19.'),
        ('b', 'La columna «Estado» se calcula sola: OK en verde a partir de '
              '75 °C, REPETIR en rojo por debajo. En regeneración también '
              'mira el tiempo: más de 60 minutos en llegar a los 75 °C es '
              'REPETIR aunque los alcance.'),
        ('h', 'Límites que aplica la hoja'),
        ('b', 'Cocción: 75 °C o más en el centro del producto.'),
        ('b', 'Regeneración: 75 °C o más en el centro Y alcanzados en menos de '
              'una hora. Es el límite crítico que declara tu análisis de '
              'peligros (fichero 12) para este PCC.'),
        ('b', 'Por qué 75 y no 65: el RD 1021/2022 —que derogó al RD '
              '3484/2000— ya no fija esas cifras; las fija tu APPCC. Este pack '
              'aplica 75 °C a los dos procesos porque es el criterio que '
              'destruye Salmonella y Listeria con margen, y porque una sola '
              'cifra no se equivoca en mitad de un servicio.'),
        ('b', 'Pescado que se vaya a consumir crudo, marinado o en salazón '
              'ligera: no se cuece, se congela. Ese control va en el fichero 18.'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Frecuencia: cada elaboración cocinada o regenerada de las que '
              'tu análisis de peligros marque como PCC. La hoja trae 40 filas.'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('16 (NUEVO): registro de cocción y regeneración con '
                   'veredicto a 75 °C —y a 60 min en regeneración— y semáforo '
                   '(DOM-01/TEC-04/COM-04/DOM-R2-07)')


# ==========================================================================
# 17 — Enfriamiento y descongelación (NUEVO, dos pestañas)
# ==========================================================================
CAB_17E = ['Fecha', 'Producto / elaboración', 'Hora inicio',
           'T.ª inicio (°C)', 'Hora a las 2 h', 'T.ª a las 2 h (°C)',
           'Tiempo (h)', 'Estado', 'Destino', 'Responsable', 'Observaciones']
ANCHOS_17E = [14, 32, 13, 14, 15, 16, 12, 14, 24, 16, 30]
CAB_17D = ['Fecha', 'Producto', 'Nº lote', 'Inicio (fecha y hora)',
           'Fin / puesta en uso (fecha y hora)', 'T.ª cámara (°C)',
           'Duración (h)', 'Estado', 'Responsable', 'Observaciones']
ANCHOS_17D = [14, 30, 18, 24, 26, 16, 14, 14, 16, 30]
F0_17, F1_17 = 5, 44


def _post_17(wb, fname, cambios):
    # ---- Enfriamiento ---------------------------------------------------
    ws = motor.hoja(wb, 'Enfriamiento')
    _titulo(ws, 'Registro de Enfriamiento Rápido', 11)
    ws['A2'] = ('Mes: _______________    De 60 °C a 10 °C o menos en un máximo '
                'de 2 horas (límite crítico del APPCC de este establecimiento; '
                'ver fichero 12).')
    motor.cabecera(ws, 4, CAB_17E, ANCHOS_17E)

    _pintar_fila(ws, F0_17, 11, calc=(7, 8), texto_izq=(2, 9, 11),
                 num=(4, 6), fmt={1: FMT_FECHA, 3: FMT_HORA, 5: FMT_HORA,
                                  4: motor.FMT_TEMP, 6: motor.FMT_TEMP,
                                  7: '0.00'})
    # MOD(E-C,1) y no E-C: un enfriamiento que empieza a las 23:30 y se mide a
    # la 01:30 daría -22 horas con la resta cruda, y la fila saldría OK por
    # comparar un negativo con 2.
    f_t = (f'=IF(NOT(AND(ISNUMBER($C{F0_17}),ISNUMBER($E{F0_17}))),"",'
           f'ROUND(MOD($E{F0_17}-$C{F0_17},1)*24,2))')
    # TEC-07 (ronda 2): la guarda exigía $D (T.ª inicio) y la fórmula NO usaba
    # $D para nada, así que una fila con desviación evidente —producto a 40 °C
    # dos horas después de empezar a enfriar— devolvía cadena vacía si el
    # operario no anotó la temperatura de partida. Como el formato condicional
    # sólo pinta texto, la celda salía BLANCA en la hoja impresa,
    # indistinguible de una fila conforme. Ahora D se USA: sin ella la fila no
    # acredita un enfriamiento y se declara INCOMPLETO (ámbar), y con ella se
    # exige además que el producto partiera de ≥60 °C.
    #
    # TEC-11: `MOD` resuelve el cruce de medianoche —para eso está— pero
    # también pliega cualquier intervalo mayor de 24 h, y un intervalo NULO
    # (misma hora en las dos casillas, que es lo que sale de copiar la celda)
    # daba 0 h y por tanto OK. Se exige $G>0.
    f_e = (f'=IF(OR($F{F0_17}="",$G{F0_17}=""),"",'
           f'IF($D{F0_17}="","INCOMPLETO",'
           f'IF(AND($D{F0_17}>=60,$F{F0_17}<=10,$G{F0_17}>0,$G{F0_17}<=2),'
           f'"OK","ALERTA")))')
    ws.cell(row=F0_17, column=7).value = f_t
    ws.cell(row=F0_17, column=8).value = f_e
    motor.reg(ws, f'G{F0_17}', f_t)
    motor.reg(ws, f'H{F0_17}', f_e)
    motor.replicar_filas(ws, F0_17, F0_17, F1_17, ncols=11, alto=20)

    dv_fecha(ws, [motor.rango('A', F0_17, F1_17)], 'Fecha',
             'Escribe una fecha (por ejemplo 05/09/2026).')
    motor.dv_decimal(ws, [motor.rango('D', F0_17, F1_17),
                          motor.rango('F', F0_17, F1_17)], 'Temperatura',
                     'Introduce la temperatura en °C, entre -30 y 130.',
                     minimo=-30, maximo=130)
    dv_hora(ws, [motor.rango('C', F0_17, F1_17),
                 motor.rango('E', F0_17, F1_17)], 'Hora',
            'Escribe la hora como HORA (por ejemplo 16:00), no como texto: la '
            'columna «Tiempo (h)» las resta y un texto no se resta.')
    motor.semaforo(ws, motor.rango('H', F0_17, F1_17))

    motor.sembrar(ws, 5, {'A': datetime.date(2026, 9, 5),
                          'B': 'Caldo de pollo (20 L)',
                          'C': datetime.time(16, 0), 'D': 92,
                          'E': datetime.time(18, 0), 'F': 6,
                          'I': 'Cámara 1 — servicio del día siguiente',
                          'J': 'A.R.'}, marca_col='K')
    motor.sembrar(ws, 6, {'A': datetime.date(2026, 9, 6),
                          'B': 'Arroz cocido (4 kg)',
                          'C': datetime.time(13, 30), 'D': 85,
                          'E': datetime.time(15, 30), 'F': 8,
                          'I': 'Cámara 2', 'J': 'M.S.'}, marca_col='K')
    motor.sembrar(ws, 7, {'A': datetime.date(2026, 9, 7),
                          'B': 'Estofado de ternera (8 kg)',
                          'C': datetime.time(17, 0), 'D': 88,
                          'E': datetime.time(19, 0), 'F': 14,
                          'I': 'Desechado', 'J': 'M.S.',
                          'K': 'Olla demasiado llena: dividido en bandejas y '
                               'desechado · incidencia INC-005'},
                  marca_col='K')

    _pie(ws, F1_17 + 2, 11, extra=[
        'Anota la hora y la temperatura al empezar y vuelve a medir a las DOS '
        'horas. Si a las 2 h el producto no ha bajado a 10 °C o menos, la '
        'columna «Estado» dice ALERTA: reparte el producto en recipientes más '
        'pequeños y planos, o deséchalo, y abre la incidencia en el '
        'registro 11.',
        'La «T.ª inicio» no es opcional: sin ella la fila dice INCOMPLETO en '
        'ámbar, porque no se puede acreditar que el producto pasara de 60 °C '
        'a 10 °C. Y si las dos horas son la misma, tampoco: un intervalo de '
        'cero no acredita nada.',
        'Truco que evita casi todas las alertas: bandejas de menos de 5 cm de '
        'altura, sin tapar hasta que baje de 20 °C, y abatidor si lo tienes.',
        'La hoja calcula el tiempo transcurrido aunque el enfriamiento cruce '
        'la medianoche.',
    ])
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    # ---- Descongelación -------------------------------------------------
    ds = motor.hoja(wb, 'Descongelación')
    _titulo(ds, 'Registro de Descongelación Controlada', 10)
    ds['A2'] = ('Mes: _______________    En refrigeración a 4 °C o menos, el '
                'producto tapado y sobre una bandeja que recoja el exudado. '
                'Nunca a temperatura ambiente.')
    motor.cabecera(ds, 4, CAB_17D, ANCHOS_17D)

    _pintar_fila(ds, F0_17, 10, calc=(7, 8), texto_izq=(2, 10),
                 num=(6,), fmt={1: FMT_FECHA, 4: FMT_FECHAHORA,
                                5: FMT_FECHAHORA, 6: motor.FMT_TEMP,
                                7: '0.0'})
    f_d = (f'=IF(NOT(AND(ISNUMBER($D{F0_17}),ISNUMBER($E{F0_17}))),"",'
           f'ROUND(($E{F0_17}-$D{F0_17})*24,1))')
    f_de = (f'=IF(OR($F{F0_17}="",$G{F0_17}=""),"",'
            f'IF(AND($F{F0_17}<=4,$G{F0_17}<=24),"OK","ALERTA"))')
    ds.cell(row=F0_17, column=7).value = f_d
    ds.cell(row=F0_17, column=8).value = f_de
    motor.reg(ds, f'G{F0_17}', f_d)
    motor.reg(ds, f'H{F0_17}', f_de)
    motor.replicar_filas(ds, F0_17, F0_17, F1_17, ncols=10, alto=20)

    dv_fecha(ds, [motor.rango('A', F0_17, F1_17)], 'Fecha',
             'Escribe una fecha (por ejemplo 05/09/2026). Si la celda se '
             'queda como texto, la hoja no puede calcular nada.')
    motor.dv_decimal(ds, [motor.rango('F', F0_17, F1_17)],
                     'Temperatura de la cámara',
                     'Introduce la temperatura de la cámara en °C, entre -5 '
                     'y 20.', minimo=-5, maximo=20)
    dv_fechahora(ds, [motor.rango('D', F0_17, F1_17),
                      motor.rango('E', F0_17, F1_17)], 'Fecha y hora',
                 'Escribe la fecha CON la hora (por ejemplo 05/09/2026 09:00). '
                 'Si la celda queda como texto, la «Duración (h)» no se puede '
                 'calcular y el Estado se queda en blanco.')
    motor.semaforo(ds, motor.rango('H', F0_17, F1_17))

    motor.sembrar(ds, 5, {'A': datetime.date(2026, 9, 5),
                          'B': 'Solomillo de cerdo congelado',
                          'C': 'L-2026-0918',
                          'D': datetime.datetime(2026, 9, 5, 9, 0),
                          'E': datetime.datetime(2026, 9, 6, 8, 0), 'F': 3,
                          'I': 'A.R.'}, marca_col='J')
    motor.sembrar(ds, 6, {'A': datetime.date(2026, 9, 7),
                          'B': 'Gambas congeladas', 'C': 'L-2026-0921',
                          'D': datetime.datetime(2026, 9, 7, 18, 0),
                          'E': datetime.datetime(2026, 9, 8, 11, 0), 'F': 2.5,
                          'I': 'M.S.'}, marca_col='J')

    _pie(ds, F1_17 + 2, 10, extra=[
        'La hoja da OK cuando la cámara está a 4 °C o menos Y el producto entra '
        'en uso dentro de las 24 horas siguientes al INICIO de la '
        'descongelación —es el mismo criterio que declara el análisis de '
        'peligros (fichero 12), y el que acota el tiempo total del producto en '
        'la zona de riesgo—. Un producto descongelado NO se vuelve a congelar.',
        'Escribe el inicio y el fin con fecha Y hora (por ejemplo '
        '05/09/2026 09:00): la columna «Duración» los resta.',
        'Descongelar a temperatura ambiente es una falta grave: aparece como '
        'tal en la guía de inspección (fichero 15) y como PCC en el análisis '
        'de peligros (fichero 12).',
    ])
    motor.IMPRESION[(fname, ds.title)] = (4, True)

    motor.escribir_instrucciones(
        wb, 'Registro de Enfriamiento y Descongelación', [
            ('h', 'Para qué sirve esta plantilla'),
            ('p', 'Cubre los dos momentos en los que un alimento pasa despacio '
                  'por la zona de peligro (entre 60 °C y 5 °C), que es donde se '
                  'multiplican las bacterias. El análisis de peligros '
                  '(fichero 12) apunta a estas dos hojas.'),
            ('h', 'Hoja «Enfriamiento»'),
            ('b', 'Regla: de 60 °C a 10 °C o menos en un máximo de 2 horas. El '
                  'RD 1021/2022 —que derogó al RD 3484/2000— ya no fija esa '
                  'cifra: la fija tu APPCC, y es la que figura como límite '
                  'crítico en el fichero 12 (Reg. (CE) 852/2004, art. 5).'),
            ('b', 'Anota hora y temperatura al empezar, y vuelve a medir a las '
                  'dos horas. El «Estado» se calcula solo: OK si el producto '
                  'partía de 60 °C o más y a las 2 h está a 10 °C o menos; '
                  'ALERTA si no; INCOMPLETO en ámbar si falta la temperatura '
                  'de partida.'),
            ('b', 'Escribe las horas como hora (16:00), no como texto: la '
                  'columna «Tiempo (h)» las resta y aguanta el cambio de día. '
                  'La celda ya no acepta texto, así que si te la rechaza es '
                  'que estabas escribiendo «16.00» o «16h».'),
            ('b', 'Si sale ALERTA: reparte en recipientes bajos o desecha, y '
                  'abre la incidencia en el registro 11.'),
            ('h', 'Hoja «Descongelación»'),
            ('b', 'Regla: en refrigeración a 4 °C o menos, tapado y sobre '
                  'bandeja, y el producto en uso dentro de las 24 horas '
                  'siguientes al INICIO de la descongelación. Es el mismo '
                  'límite crítico que figura en el fichero 12 para este PCC.'),
            ('b', 'Escribe el inicio y el fin con FECHA Y HORA. La columna '
                  '«Duración (h)» y el «Estado» se calculan solos.'),
            ('b', 'Nunca a temperatura ambiente ni bajo el grifo de agua '
                  'caliente. Un producto descongelado no se recongela.'),
            ('h', 'Frecuencia y archivo'),
            ('b', 'Frecuencia: una línea por cada enfriamiento y por cada '
                  'descongelación. 40 filas por hoja.'),
            ('b', motor.CONSERVACION),
        ], cambios)
    cambios.append('17 (NUEVO): enfriamiento (60→10 °C en 2 h) y '
                   'descongelación (≤4 °C y ≤24 h) con veredicto y semáforo')


# ==========================================================================
# 18 — Congelación preventiva de anisakis (NUEVO)
# ==========================================================================
CAB_18 = ['Fecha', 'Especie / pescado', 'Nº lote', 'Entrada (fecha y hora)',
          'Salida (fecha y hora)', 'T.ª del congelador (°C)', 'Horas',
          'Estado', 'Destino / elaboración', 'Firma', 'Observaciones']
ANCHOS_18 = [14, 26, 18, 24, 24, 20, 10, 14, 30, 14, 28]
F0_18, F1_18 = 5, 44


def _post_18(wb, fname, cambios):
    ws = motor.hoja(wb, 'Congelación Anisakis')
    _titulo(ws, 'Registro de Congelación Preventiva de Anisakis', 11)
    ws['A2'] = ('Mes: _______________    Obligatorio para todo el pescado que '
                'se sirva crudo, marinado, en escabeche o en salazón ligera.')
    motor.cabecera(ws, 4, CAB_18, ANCHOS_18)

    _pintar_fila(ws, F0_18, 11, calc=(7, 8), texto_izq=(2, 9, 11),
                 num=(6,), fmt={1: FMT_FECHA, 4: FMT_FECHAHORA,
                                5: FMT_FECHAHORA, 6: motor.FMT_TEMP,
                                7: '0.0'})
    f_h = (f'=IF(NOT(AND(ISNUMBER($D{F0_18}),ISNUMBER($E{F0_18}))),"",'
           f'ROUND(($E{F0_18}-$D{F0_18})*24,1))')
    # Las dos combinaciones legales del Reg. (CE) 853/2004, Anexo III, Secc.
    # VIII, Cap. III, D: -20 °C durante 24 h, o -35 °C durante 15 h, en TODAS
    # las partes de la pieza. La segunda no es un atajo: exige un abatidor.
    f_e = (f'=IF(OR($F{F0_18}="",$G{F0_18}=""),"",'
           f'IF(OR(AND($F{F0_18}<=-20,$G{F0_18}>=24),'
           f'AND($F{F0_18}<=-35,$G{F0_18}>=15)),"OK","ALERTA"))')
    ws.cell(row=F0_18, column=7).value = f_h
    ws.cell(row=F0_18, column=8).value = f_e
    motor.reg(ws, f'G{F0_18}', f_h)
    motor.reg(ws, f'H{F0_18}', f_e)
    motor.replicar_filas(ws, F0_18, F0_18, F1_18, ncols=11, alto=20)

    dv_fecha(ws, [motor.rango('A', F0_18, F1_18)], 'Fecha',
             'Escribe una fecha (por ejemplo 05/09/2026).')
    motor.dv_decimal(ws, [motor.rango('F', F0_18, F1_18)],
                     'Temperatura del congelador',
                     'Introduce la temperatura del congelador en °C, entre '
                     '-60 y 0.', minimo=-60, maximo=0)
    dv_fechahora(ws, [motor.rango('D', F0_18, F1_18),
                      motor.rango('E', F0_18, F1_18)], 'Fecha y hora',
                 'Escribe la fecha CON la hora (por ejemplo 05/09/2026 10:00). '
                 'Si la celda queda como texto, las «Horas» no se pueden '
                 'calcular y el Estado se queda en blanco.')
    motor.semaforo(ws, motor.rango('H', F0_18, F1_18))

    motor.sembrar(ws, 5, {'A': datetime.date(2026, 9, 5),
                          'B': 'Boquerón (Engraulis encrasicolus)',
                          'C': 'L-2026-0930',
                          'D': datetime.datetime(2026, 9, 5, 10, 0),
                          'E': datetime.datetime(2026, 9, 6, 11, 0), 'F': -22,
                          'I': 'Boquerones en vinagre', 'J': 'A.R.'},
                  marca_col='K')
    motor.sembrar(ws, 6, {'A': datetime.date(2026, 9, 9),
                          'B': 'Salmón (Salmo salar)', 'C': 'L-2026-0941',
                          'D': datetime.datetime(2026, 9, 9, 9, 0),
                          'E': datetime.datetime(2026, 9, 10, 10, 0),
                          'F': -21, 'I': 'Tartar de salmón', 'J': 'M.S.'},
                  marca_col='K')

    _pie(ws, F1_18 + 2, 11, extra=[
        'Base legal: RD 1021/2022, art. 8.1 (que derogó el RD 1420/2006) y '
        'Rgto. (CE) 853/2004, Anexo III, Secc. VIII, Cap. III.D. El '
        'tratamiento debe alcanzar -20 °C durante al menos 24 horas EN TODAS '
        'LAS PARTES de la pieza, o -35 °C durante al menos 15 horas.',
        'Un congelador doméstico de tres estrellas no llega a -20 °C con carga: '
        'mide la temperatura del equipo cada vez y anótala, porque es la '
        'columna que el inspector cruza con este registro y con el fichero 01.',
        'Se aplica a boquerones en vinagre, ceviche, tartar, sashimi, sushi, '
        'marinados, escabeches poco cocinados, salazones ligeras y arenques. '
        'NO hace falta en pescado que se cocina a 60 °C o más en el centro '
        'durante al menos un minuto, ni en el pescado de acuicultura que cumpla '
                  'la exención del Reg. (CE) 853/2004 con la acreditación '
                  'escrita del productor, '
        'ni en moluscos bivalvos.',
        'Informa al cliente: el art. 8.2 del RD 1021/2022 —el mismo que '
        'derogó el RD 1420/2006— obliga a decir, con carteles o en la '
        'carta-menú, que el pescado servido crudo o poco cocinado ha sido '
        'congelado previamente.',
    ])
    motor.IMPRESION[(fname, ws.title)] = (4, True, 'C5')  # TEC-08

    motor.escribir_instrucciones(
        wb, 'Registro de Congelación Preventiva de Anisakis', [
            ('h', 'Para qué sirve esta plantilla'),
            ('p', 'Es el registro que acredita la congelación preventiva del '
                  'pescado destinado a consumirse crudo o prácticamente crudo. '
                  'Es una obligación legal con sanción y uno de los primeros '
                  'papeles que pide el inspector en un local que sirva '
                  'boquerones en vinagre, ceviche o sushi.'),
            ('h', 'Cómo usarla'),
            ('b', 'Una línea por lote y por tratamiento. Anota la especie, el '
                  'lote y la fecha y hora de entrada y de salida del '
                  'congelador.'),
            ('b', 'Escribe entrada y salida con FECHA Y HORA (por ejemplo '
                  '05/09/2026 10:00). La columna «Horas» las resta sola.'),
            ('b', 'Mide y anota la temperatura del congelador en cada '
                  'tratamiento; el veredicto la usa.'),
            ('b', 'El «Estado» dice OK con -20 °C durante 24 h o más, o con '
                  '-35 °C durante 15 h o más. Cualquier otra combinación es '
                  'ALERTA y el pescado no se puede servir crudo.'),
            ('h', 'Cuándo es obligatorio'),
            ('b', 'Pescado que se sirva crudo o casi crudo: sashimi, sushi, '
                  'tartar, carpaccio, ceviche.'),
            ('b', 'Marinados, escabeches poco cocinados, salazones ligeras, '
                  'ahumados en frío y arenques.'),
            ('b', 'Boquerones en vinagre: el caso más frecuente de anisakiasis '
                  'en España.'),
            ('h', 'Cuándo NO hace falta'),
            ('b', 'Pescado cocinado a 60 °C o más en el centro durante al menos '
                  'un minuto (regístralo en el fichero 16).'),
            # DOM-R2-27: «criado en agua continental y alimentado con pienso»
            # no es el criterio del Reg. (CE) 853/2004: el agua continental no
            # aparece en la norma. Tal como estaba, la regla era a la vez más
            # estricta (excluía al salmón de acuicultura marina, que sí puede
            # acogerse) y más laxa (un pescado de agua dulce alimentado con
            # descartes crudos quedaba exento, y no lo está).
            ('b', 'Pescado de acuicultura criado a partir de embriones y '
                  'alimentado exclusivamente con una dieta que no pueda '
                  'contener parásitos viables, siempre que el productor lo '
                  'acredite POR ESCRITO: guarda esa ficha del proveedor, es lo '
                  'que pide el inspector. Y los moluscos bivalvos.'),
            ('h', 'Obligación de informar'),
            ('p', 'El art. 8.2 del RD 1021/2022 —el mismo que derogó el '
                  'RD 1420/2006— obliga a informar al consumidor, con '
                  'carteles o en la carta-menú, de que el pescado servido '
                  'crudo o poco cocinado ha sido congelado previamente.'),
            ('h', 'Frecuencia y archivo'),
            ('b', 'Frecuencia: cada lote de pescado que vaya a servirse crudo. '
                  'La hoja trae 40 filas.'),
            ('b', motor.CONSERVACION),
        ], cambios)
    cambios.append('18 (NUEVO): congelación preventiva de anisakis con las dos '
                   'combinaciones legales (-20 °C/24 h y -35 °C/15 h) (DOM-06)')


# ==========================================================================
# 19 — Verificación de termómetros (NUEVO)
# ==========================================================================
MET_HIELO = 'Hielo fundente (0 °C)'
MET_EBUL = 'Agua en ebullición (100 °C)'
CAB_19 = ['Fecha', 'Equipo / sonda (nº)', 'Método de verificación',
          'Referencia (°C)', 'Lectura (°C)', 'Desviación (°C)', 'Estado',
          'Acción realizada', 'Firma', 'Observaciones']
ANCHOS_19 = [14, 22, 28, 15, 14, 15, 14, 32, 14, 28]
F0_19, F1_19 = 5, 44


def _post_19(wb, fname, cambios):
    ws = motor.hoja(wb, 'Verificación Termómetros')
    _titulo(ws, 'Verificación de Termómetros y Sondas', 10)
    ws['A2'] = ('Año: _______________    Una verificación mensual por cada '
                'termómetro y sonda del local. Tolerancia: ±1 °C.')
    # DOM-R2-21 / TEC-06 (ronda 2): la referencia de la ebullición estaba
    # clavada en 100 °C y la columna no admitía corrección, así que en buena
    # parte de España una sonda IMPECABLE salía NO APTO: en Madrid (667 m) el
    # agua hierve a ~97,8 °C → desviación 2,2 contra una tolerancia de ±1, y la
    # nota del pie le decía al usuario que retirase de uso un equipo que está
    # bien. El aviso existía, pero sólo en prosa. Aquí la altitud es un dato de
    # cabecera y la referencia se calcula: 100 − altitud/300 (≈1 °C menos cada
    # 300 m). Con 0 m la hoja se comporta exactamente como antes.
    ws['A3'] = 'Altitud del local (m sobre el nivel del mar):'
    ws['A3'].font = Font(bold=True, size=11)
    motor.verde(ws.cell(row=3, column=2), motor.FMT_ENT)
    ws.cell(row=3, column=2).value = 0
    c_alt = ws.cell(row=3, column=3)
    c_alt.value = ('Corrige la referencia del método de ebullición. Déjalo en '
                   '0 si estás a nivel del mar.')
    c_alt.font = Font(size=9, italic=True)
    motor.cabecera(ws, 4, CAB_19, ANCHOS_19)

    _pintar_fila(ws, F0_19, 10, calc=(4, 6, 7), texto_izq=(3, 8, 10),
                 num=(5,), fmt={1: FMT_FECHA, 4: motor.FMT_TEMP,
                                5: motor.FMT_TEMP, 6: motor.FMT_TEMP})
    # La referencia NO se teclea: la pone el método. Un punto de referencia
    # escrito a mano es la forma más fácil de que la desviación salga 0,0 en un
    # termómetro que va mal.
    f_ref = (f'=IF($C{F0_19}="","",IF($C{F0_19}="{MET_HIELO}",0,'
             f'IF($C{F0_19}="{MET_EBUL}",'
             f'100-IF($B$3="",0,$B$3)/300,"")))')
    f_des = (f'=IF(OR($D{F0_19}="",$E{F0_19}=""),"",'
             f'ROUND(ABS($E{F0_19}-$D{F0_19}),1))')
    # TEC-12: la desviación se redondeaba a un decimal ANTES de compararla con
    # la tolerancia, así que la franja 1,00-1,04 °C —que está FUERA de ±1— se
    # declaraba APTO. El ROUND se queda sólo para mostrar la columna F; el
    # veredicto compara el valor real.
    f_est = (f'=IF($F{F0_19}="","",'
             f'IF(ABS($E{F0_19}-$D{F0_19})<=1,"APTO","NO APTO"))')
    ws.cell(row=F0_19, column=4).value = f_ref
    ws.cell(row=F0_19, column=6).value = f_des
    ws.cell(row=F0_19, column=7).value = f_est
    motor.reg(ws, f'D{F0_19}', f_ref)
    motor.reg(ws, f'F{F0_19}', f_des)
    motor.reg(ws, f'G{F0_19}', f_est)
    motor.replicar_filas(ws, F0_19, F0_19, F1_19, ncols=10, alto=20)

    dv_fecha(ws, [motor.rango('A', F0_19, F1_19)], 'Fecha',
             'Escribe una fecha (por ejemplo 01/09/2026).')
    motor.dv_lista(ws, f'"{MET_HIELO},{MET_EBUL}"',
                   [motor.rango('C', F0_19, F1_19)], 'Método de verificación',
                   'Elige uno de los dos métodos. La columna «Referencia» se '
                   'rellena sola a partir de él; si escribes texto libre, la '
                   'fila no acredita nada.')
    motor.dv_decimal(ws, [motor.rango('E', F0_19, F1_19)], 'Lectura',
                     'Introduce la lectura del termómetro en °C, entre -60 '
                     'y 150.', minimo=-60, maximo=150)
    motor.semaforo(ws, motor.rango('G', F0_19, F1_19),
                   extra_rojo=('NO APTO',))

    motor.sembrar(ws, 5, {'A': datetime.date(2026, 9, 1), 'B': 'Sonda S-01',
                          'C': MET_HIELO, 'E': 0.3, 'H': 'Ninguna',
                          'I': 'A.R.'}, marca_col='J')
    motor.sembrar(ws, 6, {'A': datetime.date(2026, 9, 1), 'B': 'Sonda S-02',
                          'C': MET_EBUL, 'E': 99.5, 'H': 'Ninguna',
                          'I': 'A.R.'}, marca_col='J')
    motor.sembrar(ws, 7, {'A': datetime.date(2026, 9, 1),
                          'B': 'Termómetro cámara 2', 'C': MET_HIELO,
                          'E': -1.8,
                          'H': 'Retirado de uso y sustituido · incidencia '
                               'INC-006', 'I': 'M.S.'}, marca_col='J')

    _pie(ws, F1_19 + 2, 10, extra=[
        'Método del hielo fundente: llena un vaso con hielo picado y añade un '
        'poco de agua fría hasta cubrirlo; remueve, espera dos minutos y mide '
        'en el centro sin tocar el fondo ni las paredes. La referencia son '
        '0 °C.',
        'Método de la ebullición: agua hirviendo a borbotones. La referencia '
        'NO son siempre 100 °C: el agua hierve aproximadamente 1 °C menos cada '
        '300 m de altitud. Escribe la altitud de tu local en la casilla verde '
        'de arriba (B3) y la columna «Referencia» se corrige sola — en Madrid '
        '(667 m) pasa a 97,8 °C. Si no la sabes, usa el hielo fundente, que no '
        'depende de la altitud.',
        'Si la desviación pasa de 1 °C, el estado dice NO APTO: recalibra el '
        'equipo si lo admite o retíralo de uso, y anótalo en la columna '
        '«Acción realizada». Un termómetro NO APTO invalida las lecturas que '
        'haya tomado desde la última verificación buena.',
        'Cada sonda debe llevar un número visible; es el que se anota en los '
        'registros 01, 02, 16, 17 y 18.',
    ])
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    motor.escribir_instrucciones(wb, 'Verificación de Termómetros y Sondas', [
        ('h', 'Para qué sirve esta plantilla'),
        ('p', 'Todo el pack decide con temperaturas, y todas las temperaturas '
              'salen de un termómetro. Si el termómetro va mal, los registros '
              'de temperatura, de cocción, de enfriamiento y de anisakis '
              'acreditan lo contrario de lo que pasó. Ésta es la hoja que '
              'demuestra que los aparatos dicen la verdad, y es uno de los 25 '
              'puntos de la guía de inspección (fichero 15).'),
        ('h', 'Cómo usarla'),
        ('b', 'Una línea por termómetro o sonda y verificación. Numera los '
              'equipos y usa siempre el mismo número en el resto de registros.'),
        ('b', 'Escribe la ALTITUD de tu local en la casilla verde de la '
              'cabecera (B3) antes de nada. Sólo afecta al método de '
              'ebullición, pero sin ella una sonda perfecta puede salir NO '
              'APTO: en Madrid, a 667 m, el agua hierve a 97,8 °C.'),
        ('b', 'Elige el método en el desplegable. La «Referencia» se rellena '
              'sola: 0 °C con hielo fundente y 100 − altitud/300 con agua en '
              'ebullición.'),
        ('b', 'Anota la lectura del aparato. La «Desviación» y el «Estado» se '
              'calculan solos: APTO en verde hasta ±1 °C, NO APTO en rojo por '
              'encima.'),
        ('h', 'Los dos métodos, paso a paso'),
        ('b', 'Hielo fundente: vaso lleno de hielo picado, agua fría justo '
              'hasta cubrirlo, remover, esperar dos minutos y medir en el '
              'centro sin tocar el vaso. Debe marcar 0 °C ±1.'),
        ('b', 'Agua en ebullición: hirviendo a borbotones, medir en el centro. '
              'Debe marcar la «Referencia» que calcula la hoja ±1 °C, que a '
              'nivel del mar son 100 °C y a 667 m, 97,8. Si no tienes clara la '
              'altitud, usa el hielo fundente.'),
        ('h', 'Qué hacer con un NO APTO'),
        ('b', 'Recalibra el equipo si tiene ajuste, o retíralo de uso y '
              'sustitúyelo.'),
        ('b', 'Revisa los registros tomados con ese equipo desde la última '
              'verificación correcta: pueden estar mal y hay que valorar si '
              'afectaron a algún producto (registro 11).'),
        ('h', 'Frecuencia y archivo'),
        ('b', 'Frecuencia recomendada: mensual para cada termómetro y sonda, y '
              'siempre después de un golpe o de una caída. La hoja trae 40 '
              'filas: un año de tres o cuatro equipos.'),
        ('b', motor.CONSERVACION),
    ], cambios)
    cambios.append('19 (NUEVO): verificación mensual de termómetros con '
                   'referencia por método corregida por altitud, tolerancia '
                   '±1 °C sin redondeo previo '
                   '(DOM-10/COM-07/DOM-R2-21/TEC-06/TEC-12)')


# ==========================================================================
# BONUS-01 — Registro de formación
# ==========================================================================
CAB_B1 = ['Nombre del empleado', 'Puesto', 'Formación recibida',
          'Fecha inicio', 'Fecha fin', 'Duración (h)', 'Entidad formadora',
          'Nº certificado', 'Válido hasta', 'Estado', 'Firma del empleado']
ANCHOS_B1 = [26, 18, 30, 14, 14, 12, 24, 18, 14, 16, 20]
F0_B1, F1_B1 = 5, 44
RES_B1 = 46
FORMACIONES_B1 = ('Higiene alimentaria (manipulación de alimentos),'
                  'APPCC básico,APPCC avanzado,Alérgenos e información al '
                  'consumidor,Limpieza y desinfección,Primeros auxilios,'
                  'Prevención de riesgos laborales,Otra')


def _post_b1(wb, fname, cambios):
    ws = motor.hoja(wb, 'Formación Personal')
    _titulo(ws, 'BONUS — Registro de Formación en Seguridad Alimentaria', 11)
    ws['A2'] = ('Establecimiento: ________________________________    '
                'Responsable: _________________________________')
    motor.cabecera(ws, 4, CAB_B1, ANCHOS_B1)

    _pintar_fila(ws, F0_B1, 11, calc=(10,), texto_izq=(1, 3, 7),
                 num=(6,), fmt={4: FMT_FECHA, 5: FMT_FECHA,
                                6: motor.FMT_ENT, 9: FMT_FECHA})
    # DOM-33/TEC-28/COM-25: el fichero cerraba exigiendo «formación acreditada
    # vigente» y era el único dato que no registraba ni calculaba. `ISNUMBER`
    # es aquí un requisito de seguridad, no un adorno: en Excel cualquier TEXTO
    # es mayor que cualquier número, así que un «01/03/2020» escrito como texto
    # daría VIGENTE a un certificado caducado hace años.
    f = (f'=IF(NOT(ISNUMBER($I{F0_B1})),"",'
         f'IF($I{F0_B1}<TODAY(),"CADUCADO",'
         f'IF($I{F0_B1}-TODAY()<60,"RENOVAR","VIGENTE")))')
    ws.cell(row=F0_B1, column=10).value = f
    motor.reg(ws, f'J{F0_B1}', f)
    motor.replicar_filas(ws, F0_B1, F0_B1, F1_B1, ncols=11, alto=20)

    motor.dv_lista(ws, f'"{FORMACIONES_B1}"',
                   [motor.rango('C', F0_B1, F1_B1)], 'Formación recibida',
                   'Elige una formación de la lista o «Otra» y detállala en el '
                   'nombre del certificado.')
    for col in 'DEI':
        dv_fecha(ws, [motor.rango(col, F0_B1, F1_B1)], 'Fecha',
                 'Escribe una fecha (por ejemplo 15/03/2026). Si la celda se '
                 'queda como texto, la columna «Estado» no puede compararla '
                 'con el día de hoy y se queda en blanco.')
    motor.dv_decimal(ws, [motor.rango('F', F0_B1, F1_B1)], 'Duración',
                     'Introduce la duración en horas, entre 0 y 500.',
                     minimo=0, maximo=500)
    motor.semaforo(ws, motor.rango('J', F0_B1, F1_B1))

    # TEC-13 (ronda 2): «Válido hasta» iba con fechas ABSOLUTAS y el Estado se
    # calcula contra TODAY(), así que el ejemplo envejecía con el cliente —en
    # una plantilla que se vende con acceso vitalicio— y ninguna de las tres
    # filas caía en RENOVAR, que es justo el estado que la v2.0 añadió y el que
    # conviene enseñar. Con fórmulas relativas a TODAY() los tres estados se
    # ven siempre, el día que el cliente abra el fichero. Las fechas de la
    # formación se derivan de la misma referencia para que la fila siga
    # leyéndose con sentido (formación → validez de 3 años).
    hoy = datetime.date.today()

    def _cert(d):
        return f'CERT-{d.year}-{d.timetuple().tm_yday:04d}'

    d5 = hoy - datetime.timedelta(days=195)      # VIGENTE (caduca en ~2,5 años)
    d6 = hoy - datetime.timedelta(days=1065)     # RENOVAR (caduca en 30 días)
    d7 = hoy - datetime.timedelta(days=1155)     # CADUCADO (venció hace 60)
    motor.sembrar(ws, 5, {'A': 'Ana Ruiz', 'B': 'Jefa de cocina',
                          'C': 'Higiene alimentaria (manipulación de '
                               'alimentos)',
                          'D': d5, 'E': d5, 'F': 6,
                          'G': 'Formación Hostelera Norte S.L.',
                          'H': _cert(d5),
                          'I': '=TODAY()+900',
                          'K': 'A.R.'}, marca_col='A')
    motor.sembrar(ws, 6, {'A': 'Marcos Sáez', 'B': 'Cocinero',
                          'C': 'Alérgenos e información al consumidor',
                          'D': d6, 'E': d6, 'F': 4,
                          'G': 'Formación Hostelera Norte S.L.',
                          'H': _cert(d6),
                          'I': '=TODAY()+30',
                          'K': 'M.S.'}, marca_col='A')
    motor.sembrar(ws, 7, {'A': 'Lucía Prat', 'B': 'Camarera',
                          'C': 'Higiene alimentaria (manipulación de '
                               'alimentos)',
                          'D': d7, 'E': d7, 'F': 6,
                          'G': 'Aula Gastro Formación',
                          'H': _cert(d7),
                          'I': '=TODAY()-60',
                          'K': 'L.P.'}, marca_col='A')

    ws.cell(row=RES_B1, column=1,
            value='Formaciones CADUCADAS:').font = Font(bold=True, size=11)
    f_cad = f'=COUNTIF($J${F0_B1}:$J${F1_B1},"CADUCADO")'
    ws.cell(row=RES_B1, column=2).value = f_cad
    motor.calculada(ws.cell(row=RES_B1, column=2), motor.FMT_ENT)
    ws.cell(row=RES_B1, column=2).font = Font(bold=True, size=12)
    motor.reg(ws, f'B{RES_B1}', f_cad)
    motor.cf_formula(ws, f'B{RES_B1}', f'=B{RES_B1}>0')

    ws.cell(row=RES_B1, column=3,
            value='Caducan en menos de 60 días:').font = Font(bold=True,
                                                              size=11)
    f_ren = f'=COUNTIF($J${F0_B1}:$J${F1_B1},"RENOVAR")'
    ws.cell(row=RES_B1, column=4).value = f_ren
    motor.calculada(ws.cell(row=RES_B1, column=4), motor.FMT_ENT)
    ws.cell(row=RES_B1, column=4).font = Font(bold=True, size=12)
    motor.reg(ws, f'D{RES_B1}', f_ren)
    motor.cf_formula(ws, f'D{RES_B1}', f'=D{RES_B1}>0',
                     bg=motor.SEM_AMB_BG, fg=motor.SEM_AMB_FG)

    _pie(ws, RES_B1 + 2, 11, extra=[
        'Escribe «Válido hasta» como FECHA. La columna «Estado» la compara con '
        'el día de hoy y responde VIGENTE, RENOVAR (quedan menos de 60 días) o '
        'CADUCADO, con semáforo. Las dos casillas de arriba cuentan cuántas hay '
        'de cada.',
        'El carné oficial de manipulador de alimentos se suprimió con el RD '
        '109/2010: lo que se exige hoy es formación en higiene alimentaria '
        'acreditada POR LA EMPRESA (Reg. (CE) 852/2004, Anexo II, Cap. XII). La '
        'norma no fija una caducidad, así que la fecha de «Válido hasta» es el '
        'plazo de reciclaje que fija tu propio plan de formación; lo habitual '
        'es entre 2 y 4 años.',
        'La firma del empleado es lo que acredita que recibió la formación, no '
        'sólo que la empresa la contrató. Sin ella, el registro vale la mitad.',
        'Las tres primeras filas son EJEMPLOS —lo dice el nombre— y enseñan a '
        'propósito los tres estados: VIGENTE, RENOVAR y CADUCADO. Su «Válido '
        'hasta» es una fórmula relativa al día de hoy, así que el ejemplo no '
        'caduca con el tiempo; al escribir tu fecha real la fórmula desaparece '
        'y la celda pasa a ser un dato normal.',
    ])
    motor.IMPRESION[(fname, ws.title)] = (4, True)

    motor.escribir_instrucciones(
        wb, 'BONUS — Registro de Formación en Seguridad Alimentaria', [
            ('h', 'Cómo usar esta plantilla'),
            ('b', 'Una línea por empleado y formación. Registra TODA la '
                  'formación en seguridad alimentaria: higiene, APPCC, '
                  'alérgenos, limpieza, primeros auxilios.'),
            ('b', 'Rellena «Válido hasta» con la fecha de reciclaje que fije tu '
                  'plan de formación. Es el dato que convierte la tabla en un '
                  'aviso: la columna «Estado» se calcula sola.'),
            ('b', 'Estados: VIGENTE (verde), RENOVAR (ámbar, quedan menos de 60 '
                  'días) y CADUCADO (rojo). Debajo de la tabla hay dos '
                  'contadores.'),
            ('b', 'Pide la FIRMA del empleado: es lo que acredita que recibió '
                  'la formación.'),
            ('b', 'Guarda una copia de cada certificado junto con este '
                  'registro. El inspector puede pedir los dos.'),
            ('h', 'Qué exige la normativa'),
            ('p', 'El carné oficial de manipulador de alimentos se suprimió con '
                  'el RD 109/2010. Desde entonces la obligación es del '
                  'operador: garantizar que quien manipula alimentos ha '
                  'recibido formación en higiene alimentaria adecuada a su '
                  'puesto y poder demostrarlo (Reg. (CE) 852/2004, Anexo II, '
                  'Cap. XII).'),
            ('b', 'La norma no fija un plazo de caducidad ni una entidad '
                  'obligatoria: lo fija tu plan de formación. Lo habitual es '
                  'reciclar cada 2-4 años y siempre que cambien los procesos.'),
            ('b', 'La formación en alérgenos es exigible desde el Reg. (UE) '
                  '1169/2011: quien atiende en sala también tiene que saber '
                  'responder.'),
            ('h', 'Nota sobre la fecha de hoy'),
            ('p', 'La columna «Estado» usa la función TODAY(), así que se '
                  'actualiza sola cada vez que abres el fichero con Excel, '
                  'LibreOffice o Google Sheets. Si lo miras en un visor que no '
                  'recalcula (la vista previa del móvil, por ejemplo), verás el '
                  'valor del día en que se generó la plantilla.'),
            ('h', 'Frecuencia y archivo'),
            ('b', 'Revisa el registro al menos una vez al trimestre y cada vez '
                  'que entre alguien nuevo. La hoja trae 40 filas.'),
            ('b', motor.CONSERVACION),
        ], cambios)
    cambios.append('BONUS-01: «Válido hasta», «Firma del empleado» y Estado '
                   'con TODAY() (VIGENTE/RENOVAR/CADUCADO) + 2 contadores '
                   '(DOM-33/TEC-28/COM-25). ⚠ TODAY() se cachea con la fecha '
                   'de generación; Excel lo recalcula al abrir.')


# ==========================================================================
# BONUS-02 — Protocolo de alerta alimentaria
# ==========================================================================
# DOM-21: el cartel mandaba llamar al 112 —que es emergencia médica— y dejaba
# el aviso a la autoridad sanitaria sin mencionar. La única obligación legal
# del operador ante la sospecha de un alimento nocivo es la del art. 19 del
# Reg. (CE) 178/2002: retirar e informar de inmediato a la autoridad
# competente, que en España es Salud Pública de la comunidad autónoma (activa
# la red SCIRI a través de AESAN).
PASOS_B2 = [
    ('PASO 1: IDENTIFICAR',
     'Identificar el producto afectado: nombre, nº de lote, proveedor y fecha '
     'de recepción.',
     'Consultar el registro 06 de trazabilidad (entrada y salida).'),
    ('PASO 2: AISLAR',
     'Retirar INMEDIATAMENTE el producto de la cadena alimentaria. No '
     'servirlo, no cocinarlo, no devolverlo al almacén general.',
     'Marcar «NO USAR — ALERTA» y separarlo en una zona aislada e '
     'identificada.'),
    ('PASO 3: NOTIFICAR A LA AUTORIDAD SANITARIA',
     'Notificar de inmediato a la autoridad sanitaria de tu comunidad autónoma '
     '(Salud Pública) y al proveedor. Es una OBLIGACIÓN LEGAL del art. 19 del '
     'Reg. (CE) 178/2002 cuando se sospecha que un alimento puesto en el '
     'mercado puede ser nocivo; la autoridad es quien activa la red de alerta '
     '(SCIRI / AESAN).',
     'El 112 SOLO si hay una persona con síntomas: es emergencia médica, no el '
     'cauce de notificación de una alerta. Avisar también al responsable de '
     'calidad o al gerente.'),
    ('PASO 4: DOCUMENTAR',
     'Registrar qué producto es, cuánto queda en stock, cuánto se ha servido y '
     'a quién.',
     'Completar el registro 11 de Acciones Correctivas con el nº de '
     'incidencia.'),
    ('PASO 5: COMUNICAR',
     'Si el producto se ha servido al público, informar a los clientes '
     'afectados que se puedan identificar y seguir las instrucciones de la '
     'autoridad sanitaria.',
     'Coordinar con el proveedor la retirada y la devolución del producto.'),
    ('PASO 6: VERIFICAR',
     'Comprobar que no queda producto afectado en ningún punto del local.',
     'Revisar cámaras, congeladores, almacén, sala, mise en place y '
     'elaboraciones en curso.'),
    ('PASO 7: REGISTRAR Y REVISAR',
     'Documentar todo el proceso: fechas, horas, acciones, personas '
     'implicadas y desenlace.',
     'Archivar con el registro 11 y revisar si hay que cambiar algún '
     'procedimiento para que no se repita.'),
]

TELEFONOS_B2 = [
    ('Salud Pública / Autoridad sanitaria de tu comunidad autónoma',
     '(completar — es la notificación obligatoria del art. 19)'),
    ('Emergencias médicas (solo si hay personas con síntomas)', '112'),
    ('Centro de salud más cercano', '(completar)'),
    ('Responsable de calidad del establecimiento', '(nombre + teléfono)'),
    ('Proveedor del producto afectado', '(nombre + teléfono)'),
    ('Empresa de control de plagas', '(nombre + teléfono)'),
]


def _post_b2(wb, fname, cambios):
    ws = motor.hoja(wb, 'Protocolo Alerta')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 46

    ws['B2'] = 'PROTOCOLO DE ACTUACIÓN ANTE ALERTA ALIMENTARIA'
    ws['B2'].font = Font(bold=True, size=18)
    ws.merge_cells('B2:C2')
    ws['B3'] = ('Seguir estos pasos EN ORDEN. No improvisar. '
                'Colgar este cartel en zona visible de cocina.')
    ws['B3'].font = Font(size=11, italic=True)
    ws.merge_cells('B3:C3')

    fila = 5
    for titulo, que, como in PASOS_B2:
        cel = ws.cell(row=fila, column=2, value=titulo)
        cel.font = Font(bold=True, size=12)
        cel.fill = PatternFill('solid', fgColor=motor.SECCION)
        ws.merge_cells(f'B{fila}:C{fila}')
        ws.row_dimensions[fila].height = 20
        for col, texto in ((2, que), (3, como)):
            c = ws.cell(row=fila + 1, column=col, value=texto)
            c.font = Font(size=10, bold=(col == 2 and 'PASO 3' in titulo))
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = motor.BORDE
        ancho = (ws.column_dimensions['B'].width
                 + ws.column_dimensions['C'].width)
        largo = max(len(que), len(como))
        ws.row_dimensions[fila + 1].height = max(
            30, 13 * (int(largo / max(20, ancho * 0.45)) + 1))
        fila += 3

    fila += 1
    cel = ws.cell(row=fila, column=2, value='TELÉFONOS Y CONTACTOS')
    cel.font = Font(bold=True, size=12)
    cel.fill = PatternFill('solid', fgColor=motor.SECCION)
    ws.merge_cells(f'B{fila}:C{fila}')
    fila += 1
    for etiqueta, valor in TELEFONOS_B2:
        c = ws.cell(row=fila, column=2, value=etiqueta)
        c.font = Font(size=10, bold=etiqueta.startswith('Salud Pública'))
        c.alignment = Alignment(wrap_text=True, vertical='center')
        c.border = motor.BORDE
        v = ws.cell(row=fila, column=3, value=valor)
        motor.verde(v, align='left')
        v.font = Font(size=10)
        ws.row_dimensions[fila].height = 24
        fila += 1

    fila += 1
    motor.nota(ws, fila,
               'Base legal: art. 19 del Reg. (CE) 178/2002 — si sospechas que '
               'un alimento que has puesto en el mercado puede ser nocivo, '
               'tienes que retirarlo e informar de inmediato a la autoridad '
               'competente. El 112 es emergencia médica y no sustituye a esa '
               'notificación.', ncols=3, bold=True)
    motor.nota(ws, fila + 1,
               'Imprime este cartel en A4 (o amplíalo a A3 desde el diálogo de '
               'impresión de tu equipo) y cuélgalo en zona visible de cocina. '
               'Todo el personal debe conocerlo.', ncols=3)
    motor.nota(ws, fila + 2, motor.CONSERVACION, ncols=3)
    motor.nota(ws, fila + 3, motor.MARCA, ncols=3)
    motor.IMPRESION[(fname, ws.title)] = (None, False, False)

    motor.escribir_instrucciones(
        wb, 'BONUS — Protocolo de Actuación ante Alerta Alimentaria', [
            ('h', 'Cómo usar esta plantilla'),
            ('b', 'Rellena los teléfonos de la tabla de contactos ANTES de '
                  'necesitarlos, empezando por el de Salud Pública de tu '
                  'comunidad autónoma.'),
            ('b', 'Imprime el cartel en A4 (o amplíalo a A3 desde el diálogo de '
                  'impresión) y cuélgalo en zona visible de cocina.'),
            ('b', 'Repasa los siete pasos con todo el equipo y déjalo anotado '
                  'en el registro de formación (BONUS-01).'),
            ('h', 'Lo que cambia respecto de lo que suele hacerse'),
            ('p', 'Ante la sospecha de un alimento nocivo, la obligación legal '
                  'no es llamar al 112: es notificar de inmediato a la '
                  'autoridad sanitaria de tu comunidad autónoma y retirar el '
                  'producto (art. 19 del Reg. (CE) 178/2002). Esa autoridad es '
                  'la que activa la red de alerta SCIRI a través de AESAN. El '
                  '112 se llama solo si hay una persona con síntomas.'),
            ('h', 'Cuándo se activa este protocolo'),
            ('b', 'Recibes una notificación de retirada de tu proveedor o de la '
                  'autoridad sanitaria.'),
            ('b', 'Detectas un producto en mal estado que ya se ha usado o '
                  'servido.'),
            ('b', 'Un cliente comunica un malestar que puede relacionarse con '
                  'algo que has servido.'),
            ('b', 'Se rompe la cadena de frío de un lote que ya se ha '
                  'distribuido o servido.'),
            ('h', 'Con qué registros se enlaza'),
            ('b', 'Registro 06 (trazabilidad): para saber de dónde vino el lote '
                  'y a qué elaboraciones y servicios fue.'),
            ('b', 'Registro 11 (acciones correctivas): ahí se documenta la '
                  'incidencia y el destino del producto no conforme.'),
            ('b', motor.CONSERVACION),
        ], cambios)
    cambios.append('BONUS-02: PASO 3 reescrito hacia la autoridad sanitaria de '
                   'la CCAA (art. 19 Reg. 178/2002), contacto de Salud Pública '
                   'en la lista, 112 solo ante síntomas y A4 con nota de '
                   'ampliación a A3 (DOM-21/DOM-30/TEC-21)')


# ==========================================================================
# API del grupo
# ==========================================================================
POST = {
    F_08: _post_08,
    F_16: _post_16,
    F_17: _post_17,
    F_18: _post_18,
    F_19: _post_19,
    F_B1: _post_b1,
    F_B2: _post_b2,
}


def crear(fname):
    """Libro NUEVO con sus pestañas vacías (los 4 registros que no existen).

    Sólo se llama en la 1.ª pasada, cuando el fichero todavía no está en la
    carpeta de trabajo; en la 2.ª (idempotencia) `main.py` ya lo encuentra en
    disco y lo carga. Por eso aquí no se pinta NADA: todo lo que define el
    aspecto lo escribe `post()`, que corre en las dos pasadas. Si algo se
    pintara aquí, la 2.ª pasada no lo volvería a poner y el gate de
    idempotencia no lo vería — dejaría de ser una prueba.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet('Instrucciones')
    for titulo in HOJAS_NUEVAS[fname]:
        wb.create_sheet(titulo)
    return wb


def pre(wb, fname, cambios):
    """No hay inserciones estructurales en el grupo C.

    El 08 y los dos bonos se reconstruyen enteros con `motor.hoja()`, que
    limpia la pestaña existente, así que no hace falta insertar columnas ni
    llevar centinelas: la idempotencia sale por reconstrucción.
    """
    return


def post(wb, fname, cambios, registro=None):
    fn = POST.get(fname)
    if fn:
        fn(wb, fname, cambios)
    if fname in TITULOS:
        p = wb.properties
        p.title = f'{TITULOS[fname]} · Pack de Plantillas APPCC'
        p.keywords = KEYWORDS
        p.category = CATEGORIA
        p.description = DESCRIPCION
        cambios.append(f'{fname}: metadata del pack (title/keywords/category)')


# ==========================================================================
# §6 — lo que hay que poder demostrar en la ronda de refutación
# ==========================================================================
def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}: {e}'


def _caso(carpeta, destino, fname, hoja, entradas, salida, esperado, lectura):
    """Escribe `entradas` en una COPIA desechable y evalúa `salida` con pycel."""
    from pycel import ExcelCompiler
    os.makedirs(destino, exist_ok=True)
    etiqueta = f"{fname.split('-')[0]}-{hoja[:6].replace(' ', '')}-{salida}"
    dst = os.path.join(destino, f'demo-c-{etiqueta}.xlsx')
    shutil.copy2(os.path.join(carpeta, fname), dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb[hoja]
    for coord, valor in entradas.items():
        ws[coord] = valor
    wb.save(dst)
    obtenido = _ev(ExcelCompiler(filename=dst), f"'{hoja}'!{salida}")
    return {
        'ref': f'{fname}:{hoja}:{salida}',
        'entradas': {f'{hoja}!{k}': (v.isoformat()
                                     if isinstance(v, (datetime.date,
                                                       datetime.time,
                                                       datetime.datetime))
                                     else v)
                     for k, v in entradas.items()},
        'esperado': esperado, 'obtenido': obtenido,
        'ok': obtenido == esperado, 'lectura': lectura,
        'copia_desechable': dst,
    }


def demos(carpeta, origen, destino):
    """Los casos del §6 que tocan al grupo C, más el barrido de derogadas."""
    casos = []

    # --- 08: la matriz ya no miente ni se queda corta --------------------
    casos.append(_caso(
        carpeta, destino, F_08, 'Matriz Alérgenos', {}, 'S9', 'Completo',
        'Paella mixta (fila 9): las 14 casillas declaradas y la especie del '
        'cereal escrita → «Completo». En v1.1 las 112 casillas de los 8 platos '
        'estaban vacías y la leyenda define el vacío como «no verificado»: la '
        'matriz que se le enseña a un comensal alérgico venía en blanco '
        '(DOM-02/TEC-03, altas).'))
    casos.append(_caso(
        carpeta, destino, F_08, 'Matriz Alérgenos',
        {'B20': 'Tarta de zanahoria'}, 'S20', SIN_VERIFICAR,
        'Plato nuevo sin declarar: «⚠ SIN VERIFICAR» en rojo. La fila avisa '
        'sola de que no se puede publicar (TEC-03).'))
    casos.append(_caso(
        carpeta, destino, F_08, 'Matriz Alérgenos',
        {'B21': 'Tarta Sacher', 'D21': 'S', 'E21': 'N', 'F21': 'S',
         'G21': 'N', 'H21': 'N', 'I21': 'N', 'J21': 'S', 'K21': 'N',
         'L21': 'N', 'M21': 'N', 'N21': 'N', 'O21': 'N', 'P21': 'N',
         'Q21': 'N'}, 'S21', FALTA_ESPECIE,
        'Las 14 casillas puestas pero el cereal sin nombrar: «⚠ FALTA '
        'ESPECIE». El Anexo II del Reg. (UE) 1169/2011 obliga a decir qué '
        'cereal, no «gluten» (DOM-20).'))
    casos.append(_caso(
        carpeta, destino, F_08, 'Matriz Alérgenos',
        {'B200': 'Plato de la fila 200', 'D200': 'S'}, 'S200', SIN_VERIFICAR,
        'La fila 200 sigue teniendo veredicto: los rangos llegan a la 205. En '
        'v1.1 la matriz dejaba de contar y de validar en la 45, mientras la '
        'landing vendía «todos los platos de tu carta» (DOM-19).'))

    # --- 16-19: los cuatro registros nuevos ------------------------------
    casos.append(_caso(
        carpeta, destino, F_16, 'Cocción y Regeneración', {'E20': 72}, 'H20',
        'REPETIR',
        'Cocción a 72 °C en el centro: REPETIR. En v1.1 este registro NO '
        'EXISTÍA y la columna J del análisis de peligros lo citaba igualmente '
        '(DOM-01/TEC-04/COM-04, altas).'))
    casos.append(_caso(
        carpeta, destino, F_16, 'Cocción y Regeneración', {'E21': 76}, 'H21',
        'OK', 'Cocción a 76 °C: OK.'))
    casos.append(_caso(
        carpeta, destino, F_17, 'Enfriamiento',
        {'C20': datetime.time(16, 0), 'D20': 60,
         'E20': datetime.time(18, 0), 'F20': 12}, 'H20', 'ALERTA',
        'De 60 °C a 12 °C en 2 horas: ALERTA. El límite crítico del APPCC es '
        'bajar a 10 °C o menos en ese plazo, y 12 °C está dentro de la zona '
        'de peligro.'))
    casos.append(_caso(
        carpeta, destino, F_17, 'Enfriamiento',
        {'C21': datetime.time(23, 30), 'D21': 70,
         'E21': datetime.time(1, 30), 'F21': 8}, 'H21', 'OK',
        'Enfriamiento que cruza la medianoche (23:30 → 01:30) y llega a 8 °C: '
        'OK. Con una resta cruda el tiempo habría salido -22 h.'))
    casos.append(_caso(
        carpeta, destino, F_17, 'Descongelación',
        {'D20': datetime.datetime(2026, 9, 5, 9, 0),
         'E20': datetime.datetime(2026, 9, 7, 9, 0), 'F20': 3}, 'H20',
        'ALERTA',
        'Descongelación de 48 h a 3 °C: ALERTA. La cámara está bien pero el '
        'producto lleva dos días descongelándose.'))
    casos.append(_caso(
        carpeta, destino, F_18, 'Congelación Anisakis',
        {'D20': datetime.datetime(2026, 9, 5, 10, 0),
         'E20': datetime.datetime(2026, 9, 6, 6, 0), 'F20': -20}, 'H20',
        'ALERTA',
        '20 horas a -20 °C: ALERTA. El Reg. (CE) 853/2004 exige 24 h a -20 °C '
        '(o 15 h a -35 °C). En v1.1 no existía ni el registro ni el peligro '
        'parasitario en todo el pack (DOM-06, alta).'))
    casos.append(_caso(
        carpeta, destino, F_18, 'Congelación Anisakis',
        {'D21': datetime.datetime(2026, 9, 5, 10, 0),
         'E21': datetime.datetime(2026, 9, 6, 10, 0), 'F21': -20}, 'H21',
        'OK', '24 horas clavadas a -20 °C: OK.'))
    casos.append(_caso(
        carpeta, destino, F_18, 'Congelación Anisakis',
        {'D22': datetime.datetime(2026, 9, 5, 10, 0),
         'E22': datetime.datetime(2026, 9, 6, 1, 0), 'F22': -36}, 'H22',
        'OK', '15 horas a -36 °C: OK (la segunda combinación legal, con '
              'abatidor).'))
    casos.append(_caso(
        carpeta, destino, F_19, 'Verificación Termómetros',
        {'C20': MET_HIELO, 'E20': 1.5}, 'G20', 'NO APTO',
        'Termómetro que marca 1,5 °C en hielo fundente: NO APTO (tolerancia '
        '±1 °C). Sin esta hoja, todas las temperaturas del pack se tomaban con '
        'sondas que nadie contrastaba (DOM-10/COM-07).'))
    casos.append(_caso(
        carpeta, destino, F_19, 'Verificación Termómetros',
        {'C21': MET_EBUL, 'E21': 99.4}, 'G21', 'APTO',
        'Sonda que marca 99,4 °C en agua hirviendo: APTO (desviación 0,6 °C).'))

    # --- BONUS-01: la formación caducada ---------------------------------
    hoy = datetime.date.today()
    casos.append(_caso(
        carpeta, destino, F_B1, 'Formación Personal',
        {'I20': hoy - datetime.timedelta(days=30)}, 'J20', 'CADUCADO',
        'Formación con «Válido hasta» hace 30 días: CADUCADO. En v1.1 el '
        'fichero exigía «formación acreditada vigente» y era el único dato que '
        'no registraba (DOM-33/TEC-28/COM-25).'))
    casos.append(_caso(
        carpeta, destino, F_B1, 'Formación Personal',
        {'I21': hoy + datetime.timedelta(days=30)}, 'J21', 'RENOVAR',
        'Caduca dentro de 30 días: RENOVAR (ámbar), con 60 días de aviso.'))
    casos.append(_caso(
        carpeta, destino, F_B1, 'Formación Personal',
        {'I22': hoy + datetime.timedelta(days=400)}, 'J22', 'VIGENTE',
        'Caduca dentro de más de un año: VIGENTE.'))
    casos.append(_caso(
        carpeta, destino, F_B1, 'Formación Personal',
        {'I23': '01/03/2020'}, 'J23', '',
        'Fecha pegada como TEXTO: el Estado se queda en blanco en vez de decir '
        'VIGENTE. En Excel cualquier texto es mayor que cualquier número, así '
        'que sin la guarda ISNUMBER un certificado caducado en 2020 saldría '
        'vigente.'))

    # --- comprobaciones que no son de fórmula ----------------------------
    estructura = []
    ws08 = openpyxl.load_workbook(os.path.join(carpeta, F_08))['Matriz Alérgenos']
    estructura.append({
        'comprobacion': 'paella con crustáceos S',
        'ref': f'{F_08}:Matriz Alérgenos:E9',
        'esperado': 'S', 'obtenido': ws08['E9'].value,
        'ok': ws08['E9'].value == 'S'})
    vacias = [f'{openpyxl.utils.get_column_letter(c)}{r}'
              for r in range(6, 14) for c in range(COL_D, COL_Q + 1)
              if ws08.cell(row=r, column=c).value not in ('S', 'T', 'N')]
    estructura.append({
        'comprobacion': 'las 112 casillas de los 8 platos de ejemplo, '
                        'declaradas (DOM-02/TEC-03)',
        'ref': f'{F_08}:Matriz Alérgenos:D6:Q13',
        'esperado': '0 casillas sin declarar',
        'obtenido': f'{len(vacias)} sin declarar' + (f' {vacias[:6]}'
                                                     if vacias else ''),
        'ok': not vacias})
    estructura.append({
        'comprobacion': 'cabecera del 8.º alérgeno completa y con ajuste de '
                        'texto (TEC-18)',
        'ref': f'{F_08}:Matriz Alérgenos:K5',
        'esperado': 'Frutos de cáscara (indicar cuál) · wrap · alto 64',
        'obtenido': f'{ws08["K5"].value!r} · '
                    f'wrap={ws08["K5"].alignment.wrap_text} · '
                    f'alto={ws08.row_dimensions[5].height}',
        'ok': (ws08['K5'].value == 'Frutos de cáscara (indicar cuál)'
               and bool(ws08['K5'].alignment.wrap_text)
               and (ws08.row_dimensions[5].height or 0) >= 40)})
    wsb2 = openpyxl.load_workbook(os.path.join(carpeta, F_B2))['Protocolo Alerta']
    textos_b2 = ' '.join(str(c.value) for row in wsb2.iter_rows() for c in row
                         if isinstance(c.value, str))
    estructura.append({
        'comprobacion': 'BONUS-02 nombra a la autoridad sanitaria y el art. 19 '
                        '(DOM-21)',
        'ref': f'{F_B2}:Protocolo Alerta',
        'esperado': 'autoridad sanitaria + art. 19 + Salud Pública',
        'obtenido': {'autoridad sanitaria': 'autoridad sanitaria' in textos_b2,
                     'art. 19': 'art. 19' in textos_b2,
                     'Salud Pública': 'Salud Pública' in textos_b2},
        'ok': all(s in textos_b2 for s in ('autoridad sanitaria', 'art. 19',
                                           'Salud Pública'))})
    estructura.append({
        'comprobacion': 'BONUS-02 en A4 con la nota de ampliación a A3 '
                        '(DOM-30/TEC-21)',
        'ref': f'{F_B2}:Protocolo Alerta:page_setup',
        'esperado': 'paperSize 9 (A4) y texto «A4 (o amplíalo a A3…)»',
        'obtenido': f'paperSize={wsb2.page_setup.paperSize} · '
                    f'nota={"A4 (o amplíalo a A3" in textos_b2}',
        'ok': (wsb2.page_setup.paperSize == 9
               and 'A4 (o amplíalo a A3' in textos_b2)})

    # §6: normas derogadas con 0 ocurrencias en los ficheros del grupo.
    normativa = []
    for fname in FICHEROS:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for etiqueta, rx in motor.PROHIBIDAS:
            donde = [f'{ws.title}!{c.coordinate}'
                     for ws in wb.worksheets for row in ws.iter_rows()
                     for c in row
                     if isinstance(c.value, str) and rx.search(c.value)]
            normativa.append({'fichero': fname, 'cadena': etiqueta,
                              'patron': rx.pattern,
                              'ocurrencias': len(donde), 'donde': donde})

    fallos = [f"§6 {c['ref']}: esperaba {c['esperado']!r}, dio {c['obtenido']!r}"
              for c in casos if not c['ok']]
    fallos += [f"§6 {e['ref']}: {e['comprobacion']} → {e['obtenido']!r}"
               for e in estructura if not e['ok']]
    fallos += [f"§6 «{n['cadena']}» sigue viva en {n['fichero']}: {n['donde']}"
               for n in normativa if n['ocurrencias']]

    return {'fallos': fallos,
            'casos_spec_6': casos,
            'comprobaciones_estructurales': estructura,
            'nota_today': ('BONUS-01!Formación Personal!J5:J44 usa TODAY(): '
                           'inject_cache congela el veredicto con la fecha de '
                           'generación (' + datetime.date.today().isoformat()
                           + '). Excel, LibreOffice y Google Sheets lo '
                           'recalculan al abrir; sólo un visor que no recalcula '
                           'enseñaría el valor viejo. Aceptado en la SPEC §4.'),
            'normativa_derogada': normativa,
            'normativa_ocurrencias_totales':
                sum(n['ocurrencias'] for n in normativa)}


# ==========================================================================
# Casos «dato FUERA de límite» que consume `main.py` (paso 5).
# ==========================================================================
_HOY = datetime.date.today()

CASOS_LIMITE = [
    {'fichero': F_08, 'hoja': 'Matriz Alérgenos',
     'entradas': {'B30': 'Bocadillo de calamares', 'D30': 'S'}, 'salida': 'S30',
     'esperado': SIN_VERIFICAR,
     'lectura': 'Plato con una sola casilla marcada de 14'},
    {'fichero': F_16, 'hoja': 'Cocción y Regeneración',
     'entradas': {'E30': 74.9}, 'salida': 'H30', 'esperado': 'REPETIR',
     'lectura': 'Cocción a 74,9 °C (una décima por debajo del límite)'},
    {'fichero': F_17, 'hoja': 'Enfriamiento',
     'entradas': {'C30': datetime.time(10, 0), 'D30': 90,
                  'E30': datetime.time(13, 0), 'F30': 9}, 'salida': 'H30',
     'esperado': 'ALERTA',
     'lectura': 'Llega a 9 °C pero tarda 3 horas (el límite es 2)'},
    {'fichero': F_17, 'hoja': 'Descongelación',
     'entradas': {'D30': datetime.datetime(2026, 9, 5, 9, 0),
                  'E30': datetime.datetime(2026, 9, 5, 21, 0), 'F30': 8},
     'salida': 'H30', 'esperado': 'ALERTA',
     'lectura': 'Descongelación de 12 h en una cámara a 8 °C'},
    {'fichero': F_18, 'hoja': 'Congelación Anisakis',
     'entradas': {'D30': datetime.datetime(2026, 9, 5, 10, 0),
                  'E30': datetime.datetime(2026, 9, 6, 12, 0), 'F30': -18},
     'salida': 'H30', 'esperado': 'ALERTA',
     'lectura': '26 h pero a -18 °C (el límite legal es -20)'},
    {'fichero': F_19, 'hoja': 'Verificación Termómetros',
     'entradas': {'C30': MET_EBUL, 'E30': 96}, 'salida': 'G30',
     'esperado': 'NO APTO', 'lectura': 'Sonda que marca 96 °C en ebullición'},
    {'fichero': F_B1, 'hoja': 'Formación Personal',
     'entradas': {'I30': _HOY - datetime.timedelta(days=1)}, 'salida': 'J30',
     'esperado': 'CADUCADO', 'lectura': 'Formación caducada ayer'},
]
