#!/usr/bin/env python3
"""
motor.py — Motor común del Pack de Plantillas APPCC v2.0.

Implementa el §1 de `pack-appcc-v2-SPEC.md`. NO toca ficheros: recibe un
`Workbook` ya cargado y lo modifica en memoria; quien guarda es `main.py`.

Qué hace (§1.1-§1.8):
  1.1  `semaforo()` — formato condicional de tres colores sobre cualquier
       columna de estado/veredicto (OK verde · VIGILAR ámbar · ALERTA rojo).
  1.2  `dv_lista()` — TODA validación con `showErrorMessage=True` y
       `errorTitle`/`error` en español. `allow_blank` sólo donde el vacío es
       legítimo (una fila de registro sin usar).
  1.3  `replicar_filas()` — extiende una fila plantilla (estilo + fórmula
       traducida + relleno verde) hasta la última fila del registro, para que
       las 31-40 filas del mes tengan todas veredicto y desplegable.
  1.4  `print_setup()` / `area_impresion()` — A4, `print_title_rows` en todas
       las hojas de registro y `print_area` que llega hasta el pie legal.
  1.5  `sembrar()` — 2-3 filas de ejemplo realistas marcadas «(ejemplo)».
  1.6  `CONSERVACION` — una sola frase de plazo de archivo en todos los pies.
  1.7  `SUSTITUCIONES` — marco normativo vigente; se aplica en `cerrar()`,
       DESPUÉS de que los grupos hayan reescrito sus textos, así que actúa
       como red de seguridad y no como parche cosmético.
  1.8  `escribir_instrucciones()` — Instrucciones reescritas por fichero, con
       la línea «Versión 2.0 · agosto 2026 · …» y la metadata de la Fase A
       conservada (creator/keywords/category intactos, `subject` a v2.0).

IDEMPOTENTE: todo lo que escribe el motor es escritura ABSOLUTA (siempre el
mismo valor en la misma celda). Los únicos pasos destructivos —insertar
columnas en 02 y 05— viven en `grupo_a.py` detrás de un centinela de cabecera.
DV y CF se VACÍAN en `aplicar()` y se reconstruyen enteras en los grupos: si se
acumulasen, la 2.ª pasada tendría el doble de reglas.

Las utilidades `_traducir_formula`, `_rangos_dv`, `_restaurar_dv`,
`_desplazar_rango`, `insertar_columna`, `print_setup` y `linea_instrucciones`
son las de `kit-escandallos-v2_0/motor.py` (referencia probada en producción el
2026-08-22); se copian aquí para no importar un módulo con rutas propias.
"""
import copy
import math
import re

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Paleta (la del pack v1.1 + los tres colores del semáforo, §1)
# ==========================================================================
VERDE = 'E8F5E9'          # celda editable
SECCION = 'E3F2FD'        # banda azul de sección
CAB = '2D2D2D'            # cabecera de tabla (texto blanco)
GRIS = 'F2F2F2'           # columna calculada auxiliar

SEM_OK_BG, SEM_OK_FG = 'C6EFCE', '006100'
SEM_AMB_BG, SEM_AMB_FG = 'FFEB9C', '9C6500'
SEM_ROJ_BG, SEM_ROJ_FG = 'FFC7CE', '9C0006'

FMT_TEMP = '0.0'
FMT_ENT = '0'
FMT_DEC2 = '0.00'

VERSION_LINE = ('Versión 2.0 · agosto 2026 · aichef.pro/pack-appcc · '
                'info@aichef.pro')
RX_VERSION = re.compile(r'^Versi[óo]n \d+\.\d+ · .*pack-appcc')
MARCA = '— Pack de Plantillas APPCC · AI Chef Pro · aichef.pro'

# §1.6 — DOM-35/TEC-17/COM-21. El pack daba tres plazos distintos (2 años en
# limpieza y recepción, 5 en trazabilidad) y le atribuía el de 5 al Reg. (CE)
# 178/2002, que NO fija ningún plazo de archivo. Una sola frase, en todos los
# pies, que dice la verdad y remite a quien sí puede fijarlo.
CONSERVACION = ('Conservar al menos 2 años (trazabilidad y proveedores: 5); '
                'el Reg. (CE) 178/2002 exige trazabilidad pero no fija plazo '
                '— consulta la guía de prácticas correctas de higiene de tu '
                'comunidad autónoma.')

# §1.7 — marco normativo vigente. Sustitución por EXPRESIÓN REGULAR, aplicada
# a toda celda de texto de todos los ficheros en `cerrar()`. Los grupos
# reescriben antes sus propios pies con la redacción buena; esto es la red que
# garantiza el «0 ocurrencias» que exige §6, incluidos los ficheros que ningún
# grupo toca.
#
# ⚠️ Por qué regex y no `str.replace` con literales: la primera versión de esta
# tabla buscaba «carné/certificado de manipulador vigente» y el pack escribe
# «Carné/certificado de manipulador de alimentos vigente» (13!'Higiene
# Personal'!B28). No casaba NADA y el gate de §6 daba verde igualmente, que es
# el peor resultado posible: un gate que certifica lo que no ha mirado.
FORMACION = ('formación en higiene alimentaria acreditada por la empresa '
             '(Reg. (CE) 852/2004, Anexo II, Cap. XII; el carné oficial se '
             'suprimió con el RD 109/2010)')
SANCIONES = ('de 5.001 a 20.000 € en las graves y hasta 600.000 € en las muy '
             'graves (Ley 17/2011)')

SUSTITUCIONES = [
    (re.compile(r'RD\s*2207/1995'),
     'Orden de 26 de enero de 1989 (Norma de Calidad de aceites y grasas '
     'calentados): máx. 25 % de compuestos polares'),
    (re.compile(r'RD\s*140/2003'), 'RD 3/2023, de 10 de enero'),
    # Sólo las formas que mencionan CARNÉ o CERTIFICADO: «Manipulador de
    # alimentos de todos los empleados» (15!C8) es un punto de checklist que
    # reescribe el grupo B con su propio texto, y meterle aquí la parrafada
    # normativa daría una frase sin sentido.
    (re.compile(r'[Cc]arn[eé]\s*(?:/\s*|\s+o\s+)?(?:certificado\s+)?'
                r'de\s+manipulador(?:\s+de\s+alimentos)?(?:\s+vigente)?',
                re.I), FORMACION),
    (re.compile(r'[Cc]ertificado\s+de\s+manipulador(?:\s+de\s+alimentos)?'
                r'(?:\s+vigente)?', re.I), FORMACION),
    (re.compile(r'GRAVE\s*/\s*MODERADA\s*/\s*LEVE'),
     'Leve / Grave / Muy grave (Ley 17/2011, arts. 50-52)'),
    # Se traga también el «de hasta» que suele precederla: si no, «sanciones de
    # hasta €60.000» quedaba como «sanciones de de 5.001 a 20.000 €».
    (re.compile(r'(?:de\s+)?(?:hasta\s+)?(?:€\s?60\.000|60\.000\s?€)'),
     SANCIONES),
]

#: Cadenas que NO pueden quedar vivas en ningún fichero PROCESADO (gate de §6).
#: También por regex, y por el mismo motivo. «GRAVE» a secas NO entra: es parte
#: de la escala correcta (Leve / Grave / Muy grave); el delator de la escala
#: inventada es «MODERADA».
PROHIBIDAS = [
    ('RD 2207/1995 (derogado por el RD 640/2006)', re.compile(r'RD\s*2207/1995')),
    ('RD 140/2003 (derogado por el RD 3/2023)', re.compile(r'RD\s*140/2003')),
    ('carné/certificado de manipulador (suprimido por el RD 109/2010)',
     re.compile(r'carn[eé]\s*(?:/|\s+o\s+)?\s*(?:certificado\s+)?de\s+'
                r'manipulador|certificado\s+de\s+manipulador', re.I)),
    ('sanción de 60.000 € (cifra inventada)',
     re.compile(r'€\s?60\.000|60\.000\s?€')),
    # DOM-R2-02 / COM-R2-01 (altas, ronda 2): el RD 3484/2000 lo derogó el RD
    # 1021/2022 (disposición derogatoria única), que además NO reproduce las
    # temperaturas fijas: remite al APPCC del operador. El pack lo citaba seis
    # veces —y en el 12 lo había ASCENDIDO al bloque «Marco normativo»—
    # mientras la FAQ de la landing citaba el RD 1021/2022. Aquí NO hay
    # sustitución automática a propósito: cambiar la sigla dejaría frases
    # falsas del tipo «el mínimo legal del RD 1021/2022 es 65 °C». Cada grupo
    # reescribe su frase; esto es el gate que impide que vuelva.
    # El lookbehind deja pasar la única mención legítima: «el RD 1021/2022,
    # que derogó al RD 3484/2000…». Al comprador de la v1.x le sirve para
    # entender por qué han cambiado las citas; lo que no puede quedar es la
    # norma derogada presentada como vigente.
    ('RD 3484/2000 (derogado por el RD 1021/2022)',
     re.compile(r'(?<!derogó al )(?<!derogó el )RD\s*3484/2000', re.I)),
    # ANIS-01/ANIS-03 (research `auditorias/guias-v2-research-sector.json`,
    # verificado en la ficha de estado del BOE): el RD 1420/2006 está DEROGADO
    # desde el 22-dic-2022 por la disposición derogatoria única.h) del RD
    # 1021/2022. La norma vigente es el art. 8.1 del RD 1021/2022 (congelación
    # a −20 °C o inferior ≥ 24 h en la totalidad del producto, o −35 °C ≥ 15 h;
    # la puede haber hecho una etapa anterior si está justificado
    # documentalmente) junto al Rgto. (CE) 853/2004, Anexo III, Secc. VIII,
    # Cap. III.D, en la redacción del Rgto. (UE) 1276/2011. La obligación de
    # informar al consumidor con cartel o carta-menú NO desaparece: pasa al
    # art. 8.2 del mismo RD 1021/2022.
    # Igual que con el 3484/2000, aquí NO hay sustitución automática: cada
    # grupo reescribe su frase entera, porque cambiar sólo la sigla dejaría el
    # apartado equivocado colgado de la nueva norma. Esto es el gate que impide
    # que la cita derogada vuelva. El lookbehind deja pasar la única mención
    # legítima: «… que derogó el RD 1420/2006».
    ('RD 1420/2006 (derogado por el RD 1021/2022)',
     re.compile(r'(?<!derogó al )(?<!derogó el )RD\s*1420/2006', re.I)),
    ('escala de gravedad inventada («MODERADA»)', re.compile(r'\bMODERADA\b')),
]

# §1.1 — vocabulario del semáforo. Los tres primeros conjuntos son los de la
# SPEC; los extras que añade cada grupo (FALTA CLORO, REPETIR, NO APTO…) se
# pasan por parámetro para no tener que tocar el motor por cada registro nuevo.
VOC_OK = ('OK', '✓', 'Cumple', 'VIGENTE', 'Completo', 'APTO')
VOC_AMBAR = ('VIGILAR', '⚠', 'INCOMPLETO', 'CADUCA PRONTO', 'RENOVAR',
             'FALTA CLORO', 'FALTA TEST')
VOC_ROJO = ('ALERTA', 'RECHAZAR', 'CAMBIAR', 'REVISAR', '✗', 'CADUCADO',
            'EXCESO')

#: Ficheros que trata el motor por su cuenta (ninguno: en este pack cada
#: fichero pertenece a un grupo, que declara su propia lista `FICHEROS`).
FICHEROS = []

#: Registro de TODA fórmula escrita: `main.py` verifica una por una que quedó
#: con valor cacheado (o que devuelve "" por diseño).
REGISTRO = []

#: (fichero, hoja) → (titulos, landscape[, freeze]). Lo rellenan los grupos en
#: `post()`; `cerrar()` lo consume para el `print_title_rows`. Se vacía en
#: `aplicar()` para que la 2.ª pasada no herede el de la anterior.
IMPRESION = {}


def reg(ws, coord, formula):
    REGISTRO.append((ws.title, coord, formula))


# ==========================================================================
# Utilidades genéricas
# ==========================================================================
RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _traducir_formula(valor, idx, eje):
    """Desplaza las referencias RELATIVAS de una fórmula al insertar fila/col."""
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        return f'{d1}{col}{d2}{fila}'

    return RX_REF.sub(_sub, valor)


def desplazar_filas(formula, delta):
    """Baja `delta` filas las referencias de fila NO ancladas ($5 no se mueve).

    Es lo que convierte la fila plantilla del registro en las 30-40 filas del
    mes sin escribir la fórmula a mano una por una (§1.3). Las referencias a
    las hojas auxiliares van ancladas ('Límites'!$A$5:$B$12) y por eso no se
    mueven: si se movieran, la fila 40 buscaría los límites 35 filas más abajo,
    donde no hay nada, y el VLOOKUP devolvería vacío en silencio.
    """
    if not (isinstance(formula, str) and formula.startswith('=')):
        return formula

    def _sub(m):
        d1, col, d2, fila = m.groups()
        if d2 == '$':
            return m.group(0)
        return f'{d1}{col}{d2}{int(fila) + delta}'

    return RX_REF.sub(_sub, formula)


def _rangos_dv(ws):
    return [({k: getattr(dv, k, None) for k in CAMPOS_DV},
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**{k: v for k, v in attrs.items() if v is not None})
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje) if idx else r)


def _desplazar_rango(ref, idx, eje):
    fuera = []
    for p in ref.split(':'):
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        fuera.append(f'{d1}{col}{d2}{fila}')
    return ':'.join(fuera)


def insertar_columna(ws, idx):
    """Inserta una columna en idx manteniendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos de columna."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos = {k: v.width for k, v in ws.column_dimensions.items() if v.width}

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra, ancho in sorted(anchos.items(),
                               key=lambda kv: -column_index_from_string(kv[0])):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = ancho


def hoja(wb, titulo, indice=None):
    """Hoja existente o nueva, SIEMPRE vacía de contenido.

    Idempotencia: la 2.ª pasada encuentra la pestaña ya creada; si no se
    limpiara, cualquier cambio de altura de la tabla dejaría restos de la
    versión anterior debajo. Excel prohíbe : \\ / ? * [ ] en los títulos, así
    que «Salida / uso interno» de la SPEC se escribe «Salida y uso interno».
    """
    if titulo in wb.sheetnames:
        ws = wb[titulo]
        # Primero DESCOMBINAR: en una celda combinada todas menos la esquina
        # son `MergedCell` y su `.value` es de sólo lectura — limpiar antes de
        # deshacer el merge revienta en la 2.ª pasada, no en la 1.ª.
        for m in list(str(r) for r in ws.merged_cells.ranges):
            ws.unmerge_cells(m)
        for row in ws.iter_rows():
            for c in row:
                c.value = None
        ws.conditional_formatting = ConditionalFormattingList()
        ws.data_validations.dataValidation = []
        return ws
    ws = wb.create_sheet(titulo) if indice is None \
        else wb.create_sheet(titulo, indice)
    return ws


# ==========================================================================
# §1.1 — Semáforo
# ==========================================================================
def _regla(colores, palabras, ancla):
    cond = ','.join(f'{ancla}="{p}"' for p in palabras)
    bg, fg = colores
    # TEC-01 (ronda 2): dentro de <formula> de un cfRule va la EXPRESIÓN, SIN
    # «=» — igual que dentro de <f> de una celda. Excel nunca emite ese «=» y
    # ante él repara o descarta la regla; openpyxl, en cambio, hace el
    # round-trip sin rechistar, así que el gate que relee con openpyxl daba
    # verde sobre las 195 reglas del semáforo entero. `gate_xml()` de main.py
    # lo comprueba ahora sobre el XML crudo.
    return FormulaRule(
        formula=[f'OR({cond})'],
        fill=PatternFill('solid', start_color=bg, end_color=bg),
        font=Font(color=fg, bold=True), stopIfTrue=True)


def semaforo(ws, rango, extra_ok=(), extra_ambar=(), extra_rojo=()):
    """Formato condicional de tres colores sobre una columna de veredicto.

    TEC-02 (alta): el pack v1.1 no tenía NI UNA regla de formato condicional
    mientras la landing vendía «alertas automáticas» y el propio fichero 01
    prometía «aparece ALERTA en rojo». Impreso —que es el uso previsto— una
    desviación era tipográficamente idéntica a un valor conforme.

    `rango` debe ser UNA sola área contigua (p. ej. 'C7:C13'): con varias
    áreas en el mismo sqref, Excel evalúa la fórmula relativa al primer
    vértice y las demás salen desplazadas. Se llama una vez por bloque.
    """
    ancla = rango.split(':')[0]
    for colores, palabras in (
            ((SEM_ROJ_BG, SEM_ROJ_FG), tuple(VOC_ROJO) + tuple(extra_rojo)),
            ((SEM_AMB_BG, SEM_AMB_FG), tuple(VOC_AMBAR) + tuple(extra_ambar)),
            ((SEM_OK_BG, SEM_OK_FG), tuple(VOC_OK) + tuple(extra_ok))):
        ws.conditional_formatting.add(rango, _regla(colores, palabras, ancla))


def cf_formula(ws, rango, formula, bg=SEM_ROJ_BG, fg=SEM_ROJ_FG, negrita=True):
    """Regla de formato condicional a medida (TEC-30: Estado=RECHAZAR con
    Aceptado=S, que es la contradicción que un inspector busca al hojear)."""
    # Los llamadores escriben la fórmula en notación de hoja («=C65>0»); aquí
    # se le quita el «=» por lo mismo que en `_regla` (TEC-01).
    ws.conditional_formatting.add(rango, FormulaRule(
        formula=[formula.lstrip('=')],
        fill=PatternFill('solid', start_color=bg, end_color=bg),
        font=Font(color=fg, bold=negrita), stopIfTrue=True))


# ==========================================================================
# §1.2 — Validación que valida
# ==========================================================================
def dv_lista(ws, formula1, rangos, titulo, error, allow_blank=True,
             prompt=None):
    """Lista desplegable que RECHAZA lo que no está en ella (TEC-10).

    Las 15 validaciones del pack v1.1 tenían todas `showErrorMessage=False`:
    el desplegable era una comodidad y Excel aceptaba en silencio cualquier
    texto tecleado a mano. Donde eso deja de ser estético es en los contadores
    que comparan por igualdad exacta: la respuesta escrita a mano no suma y
    nadie avisa.

    `allow_blank=True` es lo normal aquí: una hoja de registro se entrega con
    30-40 filas y el mes no las llena todas; el vacío es legítimo. Lo que ya
    no puede quedar suelto es la guarda de la fórmula, que es donde se corrige
    TEC-01 (E vacío coaccionado a 0).
    """
    dv = DataValidation(type='list', formula1=formula1, allow_blank=allow_blank,
                        showErrorMessage=True, errorStyle='stop',
                        errorTitle=titulo, error=error)
    if prompt:
        dv.showInputMessage = True
        dv.promptTitle = titulo
        dv.prompt = prompt
    ws.add_data_validation(dv)
    for r in rangos:
        dv.add(r)
    return dv


def dv_decimal(ws, rangos, titulo, error, minimo=None, maximo=None):
    """Validación numérica (temperaturas, % de compuestos polares…)."""
    kw = dict(type='decimal', allow_blank=True, showErrorMessage=True,
              errorStyle='stop', errorTitle=titulo, error=error)
    if minimo is not None and maximo is not None:
        kw.update(operator='between', formula1=str(minimo), formula2=str(maximo))
    elif minimo is not None:
        kw.update(operator='greaterThanOrEqual', formula1=str(minimo))
    else:
        kw.update(operator='lessThanOrEqual', formula1=str(maximo))
    dv = DataValidation(**kw)
    ws.add_data_validation(dv)
    for r in rangos:
        dv.add(r)
    return dv


# ==========================================================================
# §1.3 — Filas y rangos
# ==========================================================================
def replicar_filas(ws, plantilla, desde, hasta, ncols=None, alto=None):
    """Replica estilo + fórmulas de la fila `plantilla` en [desde, hasta].

    TEC-26/TEC-20/COM-24: ninguna hoja del pack usaba tablas estructuradas, así
    que la fila 21 de un registro de 20 quedaba muda —sin veredicto, sin
    desplegable y sin el relleno verde que la landing usa como señal de «celda
    editable»— con el mismo aspecto que una fila buena. Aquí se extiende la
    plantilla a TODAS las filas del mes de una vez.

    Sólo se copian estilo y FÓRMULAS: los valores literales de la plantilla
    (que pueden ser un ejemplo sembrado) no se propagan.
    """
    ncols = ncols or ws.max_column
    for fila in range(desde, hasta + 1):
        for col in range(1, ncols + 1):
            src = ws.cell(row=plantilla, column=col)
            dst = ws.cell(row=fila, column=col)
            dst._style = copy.copy(src._style)
            if isinstance(src.value, str) and src.value.startswith('='):
                f = desplazar_filas(src.value, fila - plantilla)
                dst.value = f
                reg(ws, dst.coordinate, f)
            elif fila != plantilla:
                dst.value = None
        if alto:
            ws.row_dimensions[fila].height = alto


def rango(col, desde, hasta):
    return f'{col}{desde}:{col}{hasta}'


# ==========================================================================
# §1.4 — Impresión
# ==========================================================================
def print_setup(ws, titulos=None, landscape=True, freeze=True):
    """`titulos` = fila de cabecera (int) o rango '1:3'. En una hoja de UN
    bloque se repite la fila de cabecera; en una de VARIOS (el 01 tiene seis,
    uno por equipo) repetir la cabecera del primero pondría «Cámara 1» encima
    de las filas del congelador en la página 2, así que allí se repite el
    encabezado del documento."""
    # TEC-08 (ronda 2): `freeze` acepta además una COORDENADA ('D6', 'C5'…).
    # Ninguna hoja del pack congelaba columnas, y el 08 mide 289 caracteres de
    # ancho: para marcar los 14 alérgenos de la fila 137 hay que irse a la
    # derecha y el nombre del plato desaparece de la pantalla. Equivocarse de
    # fila en esa hoja es una declaración de alérgenos falsa (Reg. UE
    # 1169/2011), así que ahí se fijan también Nº, plato y categoría.
    ws.page_setup.paperSize = 9                    # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.55, right=0.55, top=0.55, bottom=0.55,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if titulos:
        if isinstance(titulos, int):
            ws.print_title_rows = f'{titulos}:{titulos}'
            defecto = ws.cell(row=titulos + 1, column=1).coordinate
        else:
            ws.print_title_rows = titulos
            defecto = ws.cell(row=int(titulos.split(':')[1]) + 1,
                              column=1).coordinate
        if isinstance(freeze, str):
            ws.freeze_panes = freeze
        elif freeze:
            ws.freeze_panes = defecto


def area_impresion(ws):
    """`print_area` hasta la última celda con contenido (TEC-13/COM-23).

    04 era el único fichero con área fijada y la cortaba antes del pie: la
    hoja que el cliente archiva perdía la nota de conservación legal y la
    marca. Aquí el área SIEMPRE llega al final del contenido.
    """
    max_r = max_c = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                max_r, max_c = max(max_r, c.row), max(max_c, c.column)
    if not max_r:
        return
    ws.print_area = f'A1:{get_column_letter(max_c)}{max_r}'


# ==========================================================================
# §1.5 — Ejemplos sembrados
# ==========================================================================
def sembrar(ws, fila, valores, marca_col=None, marca='(ejemplo)'):
    """Escribe una fila de ejemplo (DOM-03/COM-09).

    La landing promete «cada plantilla viene pre-rellenada con datos reales de
    hostelería» y nueve de los diecisiete ficheros se entregaban en blanco.
    2-3 filas por registro, con fechas genéricas y marcadas «(ejemplo)» para
    que nadie las confunda con un registro real ni las archive como tal.
    """
    for col, valor in valores.items():
        idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.cell(row=fila, column=idx).value = valor
    if marca_col:
        cel = ws.cell(row=fila,
                      column=column_index_from_string(marca_col))
        actual = cel.value
        cel.value = f'{actual} {marca}' if actual else marca


# ==========================================================================
# Pintado
# ==========================================================================
BORDE = Border(*[Side(style='thin', color='D0D0D0')] * 4)


def verde(cel, fmt=None, align='center'):
    cel.fill = PatternFill('solid', fgColor=VERDE)
    cel.protection = Protection(locked=False)
    cel.border = BORDE
    if fmt:
        cel.number_format = fmt
    if align:
        cel.alignment = Alignment(horizontal=align, vertical='center',
                                  wrap_text=True)


def calculada(cel, fmt=None):
    cel.fill = PatternFill()
    cel.protection = Protection(locked=True)
    cel.border = BORDE
    cel.alignment = Alignment(horizontal='center', vertical='center',
                              wrap_text=True)
    if fmt:
        cel.number_format = fmt


def cabecera(ws, fila, textos, anchos=None, desde=1):
    for i, txt in enumerate(textos):
        cel = ws.cell(row=fila, column=desde + i, value=txt)
        cel.font = Font(bold=True, color='FFFFFF', size=11)
        cel.fill = PatternFill('solid', fgColor=CAB)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        cel.border = BORDE
        if anchos:
            ws.column_dimensions[get_column_letter(desde + i)].width = anchos[i]
    ws.row_dimensions[fila].height = 30


def banda(ws, fila, texto, ancho_cols):
    ws.cell(row=fila, column=1, value=texto)
    cel = ws.cell(row=fila, column=1)
    cel.font = Font(bold=True, size=11)
    cel.fill = PatternFill('solid', fgColor=SECCION)
    ref = f'A{fila}:{get_column_letter(ancho_cols)}{fila}'
    if ref not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells(ref)


def nota(ws, fila, texto, ncols=1, size=9, bold=False, ancho_car=None):
    """Pie de hoja LEGIBLE al imprimir.

    Sin combinar y sin ajuste de texto, una nota de 250 caracteres en A46 se
    desborda por las columnas de la derecha y el `print_area` la corta por la
    mitad — que es la mitad de TEC-13/COM-23: el registro archivado perdía la
    instrucción legal de conservación. Se combina de A hasta la última columna
    de la tabla, se activa el ajuste y se calcula el alto.
    """
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font = Font(size=size, bold=bold)
    cel.alignment = Alignment(vertical='top', wrap_text=ncols > 1)
    if ncols > 1:
        ref = f'A{fila}:{get_column_letter(ncols)}{fila}'
        if ref not in [str(r) for r in ws.merged_cells.ranges]:
            ws.merge_cells(ref)
        if ancho_car is None:
            ancho_car = sum(
                (ws.column_dimensions[get_column_letter(c)].width or 10)
                for c in range(1, ncols + 1))
        # TEC-04 (ronda 2): el ancho de columna de Excel ya está en
        # CARACTERES de la fuente por defecto (Calibri 11); estos pies van a
        # `size` pt, así que caben `11/size` veces más caracteres por unidad de
        # ancho — no 1,6. Y una línea de `size` pt necesita ~1,36·size puntos de
        # alto, no 11 fijos. Con el cálculo viejo la frase de conservación de
        # §1.6 (197 caracteres) y la escala de sanciones del 15 (351) se
        # imprimían cortadas: openpyxl marca customHeight, así que Excel NO
        # autoajusta y lo que sobra no se ve ni en pantalla ni en papel.
        cpl = max(20, ancho_car * 11.0 / size)
        lineas = max(1, math.ceil(len(texto) / cpl))
        ws.row_dimensions[fila].height = max(12, round(lineas * size * 1.36, 1))


# ==========================================================================
# §1.8 — Instrucciones
# ==========================================================================
def escribir_instrucciones(wb, titulo, bloques, cambios):
    """Reescribe la hoja Instrucciones entera (§1.8).

    `bloques` es una lista de tuplas ('h'|'p'|'b', texto): encabezado de
    sección, párrafo suelto o viñeta '▸'. El motor añade siempre la marca y la
    línea de versión al final, así que ni el título ni la versión se repiten.
    """
    if 'Instrucciones' not in wb.sheetnames:
        return
    ws = wb['Instrucciones']
    for row in ws.iter_rows():
        for c in row:
            c.value = None
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 95

    cel = ws.cell(row=2, column=2, value=titulo)
    cel.font = Font(bold=True, size=18)
    cel.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 26

    fila = 4
    for tipo, texto in bloques:
        if tipo == 'h':
            fila += 0 if fila == 4 else 1
            c = ws.cell(row=fila, column=2, value=texto)
            c.font = Font(bold=True, size=12)
            fila += 2
        else:
            c = ws.cell(row=fila, column=2,
                        value=('▸ ' + texto) if tipo == 'b' else texto)
            c.font = Font(size=11)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            fila += 1
    fila += 1
    c = ws.cell(row=fila, column=2, value=MARCA)
    c.font = Font(bold=True, size=12)
    fila += 2
    c = ws.cell(row=fila, column=2, value=VERSION_LINE)
    c.font = Font(bold=True, size=12)
    cambios.append(f'Instrucciones reescritas ({len(bloques)} bloques) '
                   f'+ línea de versión 2.0')
    print_setup(ws, None, landscape=False, freeze=False)
    return ws


def linea_instrucciones(ws, texto, rx=None):
    """Sustituye la línea de versión (o la añade). Nunca duplica."""
    col = 2 if ws.cell(row=2, column=2).value else 1
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)


# ==========================================================================
# §1.7 — Sustituciones normativas
# ==========================================================================
def sustituir_normativa(wb, cambios):
    hechas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str) or c.value.startswith('='):
                    continue
                nuevo = c.value
                for rx, bueno in SUSTITUCIONES:
                    nuevo = rx.sub(lambda _m, b=bueno: b, nuevo)
                if nuevo != c.value:
                    hechas.append(f'{ws.title}!{c.coordinate}')
                    c.value = nuevo
    if hechas:
        cambios.append(f'§1.7 normativa vigente en {len(hechas)} celdas: '
                       + ', '.join(hechas[:6])
                       + ('…' if len(hechas) > 6 else ''))
    return hechas


def restos_prohibidos(wb):
    fuera = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str):
                    continue
                for etiqueta, rx in PROHIBIDAS:
                    if rx.search(c.value):
                        fuera.append(f'{ws.title}!{c.coordinate}: {etiqueta} '
                                     f'→ {c.value[:80]!r}')
    return fuera


# ==========================================================================
# Metadata (la de la Fase A se CONSERVA; sólo sube la versión)
# ==========================================================================
def set_metadata(wb, cambios):
    p = wb.properties
    if p.subject and 'v1.1' in p.subject:
        p.subject = p.subject.replace('v1.1', 'v2.0')
        cambios.append(f'metadata subject → {p.subject}')
    elif not p.subject:
        p.subject = 'Pack de Plantillas APPCC · v2.0'
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'


# ==========================================================================
# API principal
# ==========================================================================
def aplicar(wb, fname, cambios):
    """Paso previo común: deja el libro limpio de DV y CF para que los grupos
    los reconstruyan enteros.

    Sin esto la 2.ª pasada de la prueba de idempotencia acumularía las reglas:
    el mismo fichero con el doble de formato condicional es indistinguible a
    ojo y rompe el gate.
    """
    IMPRESION.clear()
    n_dv = n_cf = 0
    for ws in wb.worksheets:
        n_dv += len(ws.data_validations.dataValidation)
        n_cf += len(list(ws.conditional_formatting))
        ws.data_validations.dataValidation = []
        ws.conditional_formatting = ConditionalFormattingList()
    cambios.append(f'reset de {n_dv} DV y {n_cf} bloques CF previos '
                   '(se reconstruyen enteros)')


def cerrar(wb, fname, cambios):
    """Cierre común: §1.7, limpieza de '', impresión A4 y metadata."""
    sustituir_normativa(wb, cambios)
    for ws in wb.worksheets:
        # TEC-09: los 17 ficheros heredados traían `<mergeCell ref="B2"/>` —un
        # rango combinado de UNA celda, que Excel no emite nunca— y los cuatro
        # nuevos no, así que además era una incoherencia dentro del mismo pack.
        # Deshacerlos aquí es lo único que hace falta: no combinan nada.
        for m in [str(r) for r in ws.merged_cells.ranges if ':' not in str(r)]:
            ws.unmerge_cells(m)
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    c.value = None                 # censo: empty_str = 0
        if ws.title == 'Instrucciones':
            ws.page_setup.paperSize = 9
            linea_instrucciones(ws, VERSION_LINE, RX_VERSION)
            area_impresion(ws)
            continue
        cfg = IMPRESION.get((fname, ws.title), (None, True, True))
        hr, land, frz = (cfg + (True,))[:3] if len(cfg) == 2 else cfg
        print_setup(ws, hr, landscape=land, freeze=frz)
        area_impresion(ws)
    set_metadata(wb, cambios)
