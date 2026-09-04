#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_auditoria-interna-servicio.py — Libro 7 del pack «Manual del Manager de
Restaurante» (SPEC §2.2, fila 7).

Genera `build/auditoria-interna-servicio.xlsx`:

  Instrucciones · Auditoría · Resumen por Área · Histórico

60 puntos de control en 6 áreas, con peso de 1 a 3 y tres visitas. Mide la
EXPERIENCIA del cliente y el estándar de marca, que es lo que un comensal ve y
lo que ningún registro sanitario recoge.

EXCLUYE A PROPÓSITO APPCC Y SEGURIDAD ALIMENTARIA. Mezclarlas produciría un
checklist que no sirve ni para una cosa ni para la otra: los registros
sanitarios son el Pack APPCC, con su propia trazabilidad y sus propias firmas.
La hoja «Instrucciones» lo dice en voz alta.

DECISIONES TÉCNICAS
-------------------
* La puntuación de cada visita es PONDERADA por el peso del punto:
  `SUMPRODUCT(peso,puntuación)/SUMPRODUCT(peso,--ISNUMBER(puntuación))`. El
  denominador descuenta los puntos NO valorados, que es distinto de puntuarlos
  con un 0.
* La escala máxima, los dos umbrales del semáforo y la tolerancia con la que
  se lee la tendencia viven en celdas verdes: cero constantes dentro de una
  fórmula.
* Los cuatro parámetros se ESPEJAN en «Resumen por Área» porque el formato
  condicional de Excel no puede referirse a otra hoja.
* Funciones prohibidas (INDIRECT, COUNTA, PMT, OFFSET, XLOOKUP, LET, LAMBDA,
  matrices dinámicas y RANK, que pycel no implementa): cero.
* «Sin dato» = `""`, nunca `0`; `IFERROR(...,"")` en todo cociente.

Salida fija (sin argumentos):
`<carpeta>/build/auditoria-interna-servicio.xlsx`
"""
import json
import os
import sys

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/'
       'guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

PRODUCTO = 'manual-manager-restaurante'
motor.CTX['producto'] = PRODUCTO

# --------------------------------------------------------------------------
NOMBRE = 'auditoria-interna-servicio'
TITULO_LIBRO = 'Auditoría Interna de Servicio'
SUBTITULO = 'AI Chef Pro · aichef.pro — Manual del Manager de Restaurante'
SUBJECT = 'Manual del Manager de Restaurante · v1.0 · septiembre 2026'
VERSION = DE.VERSION_LINE
BIO = DE.BIO
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde: '
               'Revisar, Desproteger hoja (no tiene contraseña).')
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL, GRIS_BG = 'FFF8E1', '1565C0', 'F2F2F2'
PCT, ENT = motor.FMT_PCT, motor.FMT_ENT
FECHA = motor.FMT_FECHA
DEC2 = '#,##0.00'

# --- filas de la hoja «Auditoría» -----------------------------------------
AU_FECHA = 3
AU_AUDITOR = 4
AU_CAB = 6
AU_INI = 7
AU_FIN = AU_INI + len(DE.AUDITORIA) - 1
AU_VAL = AU_FIN + 2
AU_PESO = AU_VAL + 1
AU_SUMA = AU_VAL + 2
AU_POND = AU_VAL + 3
AU_PCT = AU_VAL + 4
AU_SIMPLE = AU_VAL + 5
AU_ESCALA = AU_VAL + 7
AU_UM_V = AU_ESCALA + 1
AU_UM_A = AU_ESCALA + 2
AU_TOL = AU_ESCALA + 3
COLS_VISITA = ['E', 'G', 'I']            # puntuación de cada visita
COLS_OBS = ['F', 'H', 'J']               # observación de cada visita

# --- filas de la hoja «Resumen por Área» ----------------------------------
RA_ESCALA = 4
RA_UM_V = 5
RA_UM_A = 6
RA_TOL = 7
RA_CAB = 9
RA_INI = 10
RA_FIN = RA_INI + len(DE.AREAS_AUDITORIA) - 1
RA_TOT = RA_FIN + 1
COLS_RA = ['D', 'E', 'F']                # visita 1, 2 y 3

# --- filas de la hoja «Histórico» -----------------------------------------
HI_UM_V = 3                              # espejo del umbral verde
HI_UM_A = 4                              # espejo del umbral ámbar
HI_CAB = 5
HI_INI = 6
HI_FIN = HI_INI + len(DE.AUDITORIAS_HECHAS) - 1
HI_RES = HI_FIN + 2
COLS_HI_AREA = ['F', 'G', 'H', 'I', 'J', 'K']

AUD = 'Auditoría'
RES = "'Resumen por Área'"


# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto, wrap=True)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9                      # A4
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezados(ws, fila, cols, alto=40):
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12, color=GOLD)


def cf(ws, rango, formula, bg, fg):
    """Regla de formato condicional SIN purgar las anteriores del mismo rango.

    `motor.regla_expresion` limpia las reglas cuyo sqref coincide, así que
    llamarla tres veces sobre el mismo rango dejaría sólo la última: el
    semáforo de tres niveles necesita apilar verde, ámbar y rojo.
    """
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


def semaforo_3(ws, rango, ancla, verde_ref, ambar_ref):
    """Verde si alcanza el umbral alto, ámbar si alcanza el bajo, rojo si no.

    Las tres reglas llevan `ISNUMBER`: sin él, una celda con `""` se compara
    como texto y el rojo caería justo sobre la que dice que no hay dato.
    """
    cf(ws, rango, '=AND(ISNUMBER({a}),{a}>={v})'.format(a=ancla, v=verde_ref),
       motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    cf(ws, rango, '=AND(ISNUMBER({a}),{a}>={m})'.format(a=ancla, m=ambar_ref),
       motor.CF_AMBAR_BG, motor.CF_AMBAR_FG)
    cf(ws, rango, '=AND(ISNUMBER({a}),{a}<{m})'.format(a=ancla, m=ambar_ref),
       motor.CF_ROJO_BG, motor.CF_ROJO_FG)


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. Imprime la hoja «Auditoría» o llévala en el móvil y recorre el local '
    'como si fueras un cliente: desde la reserva hasta la despedida en la '
    'puerta. Puntúa cada punto de 0 a 5 en la columna de la visita.',
    '2. Escribe la observación al lado. Un 2 sin observación no sirve de nada '
    'dentro de un mes: nadie se acuerda de por qué era un 2.',
    '3. Pon la fecha y quién audita en la cabecera de cada visita. Alterna '
    'visitas propias con alguna de cliente misterioso: uno ve lo que conoce, '
    'el otro ve lo que un cliente ve de verdad.',
    '4. El peso de cada punto (1 a 3) es tuyo. Un aseo sucio no vale lo mismo '
    'que una planta seca, y por eso la puntuación es PONDERADA.',
    '5. Si un punto no aplica en tu casa, deja la celda VACÍA: su peso sale '
    'del cálculo. Puntuar 0 es decir que se hace mal, no que no aplica.',
    '6. Lee «Resumen por Área». Ahí está lo que la media global esconde: un '
    'área puede empeorar visita tras visita mientras el total sube.',
    '7. «Histórico» pone las tres visitas una debajo de otra y las dibuja. '
    'Una auditoría suelta es una foto; tres seguidas son una tendencia, que '
    'es lo único que sirve para decidir.',
    '8. Cada punto por debajo del umbral ámbar se convierte en una línea del '
    'plan de acción del libro de reuniones, con responsable y fecha. Auditar '
    'sin plan de acción es hacerse una lista de quejas a uno mismo.',
]

NOTAS = [
    'ESTE LIBRO EXCLUYE A PROPÓSITO EL APPCC Y LA SEGURIDAD ALIMENTARIA. '
    'Aquí no hay temperaturas de cámara, ni registros de recepción, ni '
    'trazabilidad, ni limpieza de superficies de manipulación. Eso es el PACK '
    'APPCC de AI Chef Pro, que trae sus registros, sus frecuencias y sus '
    'firmas. Mezclar las dos cosas produce un checklist que no vale ni para '
    'auditar el servicio ni para pasar una inspección de Sanidad.',
    'Lo que sí mide esta hoja es la EXPERIENCIA y el estándar de marca: '
    'llegada y reserva, sala y ambiente, servicio y tiempos, producto y '
    'presentación, aseos y limpieza de cara al cliente, y marca y digital. '
    'Seis áreas, 60 puntos.',
    'La limpieza que aparece aquí es la que ve el cliente (aseos, mantelería, '
    'cristalería, suelo de sala). La limpieza de las superficies donde se '
    'manipula alimento es APPCC y no está en este libro.',
    'Las tres visitas sembradas son un EJEMPLO del restaurante modelado del '
    'pack. Fíjate en lo que enseñan: el total mejora visita tras visita y, al '
    'mismo tiempo, «Aseos y limpieza» empeora. Esa es exactamente la razón de '
    'que exista la hoja «Resumen por Área».',
    'La puntuación es PONDERADA por el peso de cada punto, y la media simple '
    'está al lado para contrastar. Cuando las dos se separan, es que estás '
    'fallando justo en lo que más pesa.',
    'Los umbrales del semáforo y la tolerancia con la que se lee la tendencia '
    'están en celdas verdes al pie de la hoja «Auditoría». Cámbialos y todo '
    'el libro se recalcula.',
    'Tres visitas caben en este libro. Para el trimestre siguiente, guarda una '
    'copia con otro nombre: así conservas el histórico sin machacarlo.',
    'La columna «Área» no está en verde a propósito: es el criterio con el que '
    'el libro agrupa los 60 puntos y con el que rotula el histórico. Los '
    'puntos de control sí se editan uno a uno. Si necesitas otras áreas, '
    'desprotege las hojas y cámbialas en los tres sitios: «Auditoría», '
    '«Resumen por Área» y las cabeceras de «Histórico».',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Auditoría»
# --------------------------------------------------------------------------
def hoja_auditoria(wb):
    ws = wb.create_sheet('Auditoría')
    cabecera(ws, 'Auditoría de servicio — 60 puntos, 6 áreas, 3 visitas')

    motor.val(ws, 'C%d' % AU_FECHA, 'Fecha de la visita', bold=True)
    motor.val(ws, 'C%d' % AU_AUDITOR, 'Auditor', bold=True)
    for j, col in enumerate(COLS_VISITA):
        visita = DE.AUDITORIAS_HECHAS[j] if j < len(DE.AUDITORIAS_HECHAS) \
            else None
        if visita:
            motor.val(ws, '%s%d' % (col, AU_FECHA), DE._fecha(visita[1]),
                      fmt=FECHA, verde_=True)
            motor.val(ws, '%s%d' % (col, AU_AUDITOR), visita[2], verde_=True)
        else:
            ws['%s%d' % (col, AU_FECHA)].number_format = FECHA
            motor.verde(ws, '%s%d' % (col, AU_FECHA))
            motor.verde(ws, '%s%d' % (col, AU_AUDITOR))
    motor.dv_fecha(ws, ['%s%d' % (c, AU_FECHA) for c in COLS_VISITA])

    encabezados(ws, AU_CAB, [
        ('A', '#', 5), ('B', 'Área', 24), ('C', 'Punto de control', 66),
        ('D', 'Peso (1-3)', 9),
        ('E', 'Visita 1: puntuación (0-5)', 12),
        ('F', 'Visita 1: observación', 30),
        ('G', 'Visita 2: puntuación (0-5)', 12),
        ('H', 'Visita 2: observación', 30),
        ('I', 'Visita 3: puntuación (0-5)', 12),
        ('J', 'Visita 3: observación', 30),
    ])
    ws.freeze_panes = 'D7'

    v_peso, v_punt = [], []
    for i, (num, area, punto, peso) in enumerate(DE.AUDITORIA):
        r = AU_INI + i
        motor.val(ws, 'A%d' % r, num, fmt=ENT)
        motor.val(ws, 'B%d' % r, area)
        motor.val(ws, 'C%d' % r, punto, wrap=True)
        motor.val(ws, 'D%d' % r, peso, fmt=ENT, verde_=True)
        v_peso.append('D%d' % r)
        for j, col in enumerate(COLS_VISITA):
            if j < len(DE.AUDITORIAS_HECHAS):
                motor.val(ws, '%s%d' % (col, r),
                          DE.AUDITORIAS_HECHAS[j][3][i], fmt=ENT)
            else:
                ws['%s%d' % (col, r)].number_format = ENT
            v_punt.append('%s%d' % (col, r))
        # El «Punto de control» sí se edita; el «Área» NO va en verde a
        # propósito: es el criterio con el que «Resumen por Área» agrupa con
        # SUMPRODUCT y con el que «Histórico» rotula sus columnas. Cambiarla
        # aquí sin cambiarla allí dejaría el resumen en blanco sin un solo
        # aviso. Quien necesite otras áreas desprotege la hoja y las cambia en
        # los tres sitios.
        motor.verde(ws, 'C%d' % r)
        for col in COLS_VISITA + COLS_OBS:
            motor.verde(ws, '%s%d' % (col, r))
    motor.dv_numerica(ws, v_peso, minimo=1, maximo=3, titulo='Peso (1-3)',
                      mensaje='1 conviene, 2 importa, 3 es innegociable en tu '
                              'estándar.')
    motor.dv_numerica(ws, v_punt, minimo=0, maximo=5,
                      titulo='Puntuación (0-5)',
                      mensaje='Puntúa de 0 a 5. Si el punto NO aplica en tu '
                              'casa, deja la celda vacía: su peso sale del '
                              'cálculo.')
    semaforo_3(ws, '{a}{i}:{b}{f}'.format(a=COLS_VISITA[0], i=AU_INI,
                                          b=COLS_VISITA[-1], f=AU_FIN),
               '$E%d' % AU_INI, '$D$%d' % AU_UM_V, '$D$%d' % AU_UM_A)

    seccion(ws, 'A%d' % (AU_VAL - 1),
            'RESULTADO POR VISITA — lo calcula el libro')
    motor.val(ws, 'C%d' % AU_VAL, 'Puntos valorados')
    motor.val(ws, 'C%d' % AU_PESO, 'Peso valorado')
    motor.val(ws, 'C%d' % AU_SUMA, 'Suma de peso por puntuación')
    motor.val(ws, 'C%d' % AU_POND, 'PUNTUACIÓN PONDERADA (0-5)', bold=True)
    motor.val(ws, 'C%d' % AU_PCT, 'CUMPLIMIENTO (%)', bold=True)
    motor.val(ws, 'C%d' % AU_SIMPLE, 'Media simple (0-5)')
    for col in COLS_VISITA:
        motor.f(ws, '%s%d' % (col, AU_VAL),
                '=SUMPRODUCT(--ISNUMBER({c}${a}:{c}${b}))'
                .format(c=col, a=AU_INI, b=AU_FIN), fmt=ENT)
        motor.f(ws, '%s%d' % (col, AU_PESO),
                '=SUMPRODUCT($D${a}:$D${b},--ISNUMBER({c}${a}:{c}${b}))'
                .format(c=col, a=AU_INI, b=AU_FIN), fmt=ENT)
        motor.f(ws, '%s%d' % (col, AU_SUMA),
                '=IFERROR(SUMPRODUCT($D${a}:$D${b},{c}${a}:{c}${b}),"")'
                .format(c=col, a=AU_INI, b=AU_FIN), fmt=ENT)
        cel = motor.f(ws, '%s%d' % (col, AU_POND),
                      '=IFERROR(IF({c}${p}=0,"",{c}${s}/{c}${p}),"")'
                      .format(c=col, p=AU_PESO, s=AU_SUMA), fmt=DEC2,
                      bold=True)
        cel.fill = PatternFill('solid', fgColor=CREMA)
        cel = motor.f(ws, '%s%d' % (col, AU_PCT),
                      '=IFERROR(IF(OR({c}${o}="",$D${e}=""),"",'
                      '{c}${o}/$D${e}),"")'
                      .format(c=col, o=AU_POND, e=AU_ESCALA), fmt=PCT,
                      bold=True)
        cel.fill = PatternFill('solid', fgColor=CREMA)
        motor.f(ws, '%s%d' % (col, AU_SIMPLE),
                '=IFERROR(IF({c}${v}=0,"",AVERAGE({c}${a}:{c}${b})),"")'
                .format(c=col, v=AU_VAL, a=AU_INI, b=AU_FIN), fmt=DEC2)
    semaforo_3(ws, '{a}{r}:{b}{r}'.format(a=COLS_VISITA[0], b=COLS_VISITA[-1],
                                          r=AU_POND),
               'E$%d' % AU_POND, '$D$%d' % AU_UM_V, '$D$%d' % AU_UM_A)

    seccion(ws, 'A%d' % (AU_ESCALA - 1),
            'PARÁMETROS DEL LIBRO — celdas verdes')
    motor.val(ws, 'C%d' % AU_ESCALA, 'Escala máxima de puntuación')
    motor.val(ws, 'D%d' % AU_ESCALA, 5, fmt=ENT, verde_=True)
    motor.val(ws, 'F%d' % AU_ESCALA,
              'Con qué nota máxima puntúas. De aquí sale el porcentaje de '
              'cumplimiento.', wrap=True)
    motor.val(ws, 'C%d' % AU_UM_V, 'Umbral verde (0-5)')
    motor.val(ws, 'D%d' % AU_UM_V, 4.00, fmt=DEC2, verde_=True)
    motor.val(ws, 'F%d' % AU_UM_V,
              'A partir de esta puntuación, la celda se pone en verde: es tu '
              'estándar cumplido.', wrap=True)
    motor.val(ws, 'C%d' % AU_UM_A, 'Umbral ámbar (0-5)')
    motor.val(ws, 'D%d' % AU_UM_A, 3.00, fmt=DEC2, verde_=True)
    motor.val(ws, 'F%d' % AU_UM_A,
              'Por debajo de esta puntuación, rojo: no es un matiz, es algo '
              'que hay que arreglar antes de la próxima visita.', wrap=True)
    motor.val(ws, 'C%d' % AU_TOL,
              'Tolerancia para leer la tendencia (puntos)')
    motor.val(ws, 'D%d' % AU_TOL, 0.05, fmt=DEC2, verde_=True)
    motor.val(ws, 'F%d' % AU_TOL,
              'Por debajo de esta diferencia entre la primera visita y la '
              'última, el área se lee como «Igual»: no toda oscilación es una '
              'tendencia.', wrap=True)
    motor.dv_numerica(ws, ['D%d' % AU_ESCALA], minimo=1, maximo=10,
                      titulo='Escala máxima',
                      mensaje='Escribe la nota máxima de tu escala (por '
                              'defecto 5).')
    motor.dv_numerica(ws, ['D%d' % AU_UM_V, 'D%d' % AU_UM_A, 'D%d' % AU_TOL],
                      minimo=0, maximo=10, titulo='Umbral o tolerancia',
                      mensaje='Escribe un valor de la misma escala que las '
                              'puntuaciones.')

    motor.val(ws, 'C%d' % (AU_TOL + 2),
              'Un punto con peso 3 y puntuación 2 pesa seis veces más en el '
              'resultado que un punto con peso 1 y la misma nota. Por eso los '
              'pesos se ponen ANTES de auditar, no después de ver el '
              'resultado.', wrap=True)
    setup(ws, titulos='$%d:$%d' % (AU_CAB, AU_CAB))
    return ws


# --------------------------------------------------------------------------
# Hoja «Resumen por Área»
# --------------------------------------------------------------------------
def hoja_resumen_area(wb):
    ws = wb.create_sheet('Resumen por Área')
    cabecera(ws, 'Resumen por área — lo que la media global esconde')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    for col in COLS_RA:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 46

    seccion(ws, 'A3', 'PARÁMETROS EN VIGOR — se leen de la hoja «Auditoría»; '
                      'cámbialos allí, no aquí')
    for fila, etiqueta, origen, fmt in (
            (RA_ESCALA, 'Escala máxima de puntuación', AU_ESCALA, ENT),
            (RA_UM_V, 'Umbral verde (0-5)', AU_UM_V, DEC2),
            (RA_UM_A, 'Umbral ámbar (0-5)', AU_UM_A, DEC2),
            (RA_TOL, 'Tolerancia para leer la tendencia (puntos)', AU_TOL,
             DEC2)):
        motor.val(ws, 'A%d' % fila, etiqueta)
        motor.f(ws, 'D%d' % fila,
                '=IF({s}!$D${o}="","",{s}!$D${o})'.format(s=AUD, o=origen),
                fmt=fmt)
        ws['D%d' % fila].fill = PatternFill('solid', fgColor=GRIS_BG)

    encabezados(ws, RA_CAB, [
        ('A', 'Área', None), ('B', 'Puntos de control', None),
        ('C', 'Peso del área', None), ('D', 'Visita 1 (0-5)', None),
        ('E', 'Visita 2 (0-5)', None), ('F', 'Visita 3 (0-5)', None),
        ('G', 'Variación de la visita 1 a la 3', None),
        ('H', 'Tendencia', None), ('I', 'Cumplimiento visita 3 (%)', None),
        ('J', 'Notas', None),
    ])

    for i, area in enumerate(DE.AREAS_AUDITORIA):
        r = RA_INI + i
        motor.val(ws, 'A%d' % r, area)
        motor.f(ws, 'B%d' % r,
                '=COUNTIF({s}!$B${a}:$B${b},$A{r})'.format(s=AUD, a=AU_INI,
                                                           b=AU_FIN, r=r),
                fmt=ENT)
        motor.f(ws, 'C%d' % r,
                '=IFERROR(SUMPRODUCT(--({s}!$B${a}:$B${b}=$A{r}),'
                '{s}!$D${a}:$D${b}),"")'.format(s=AUD, a=AU_INI, b=AU_FIN,
                                                r=r), fmt=ENT)
        for j, col in enumerate(COLS_RA):
            vis = COLS_VISITA[j]
            motor.f(ws, '%s%d' % (col, r),
                    '=IFERROR(SUMPRODUCT(--({s}!$B${a}:$B${b}=$A{r}),'
                    '{s}!$D${a}:$D${b},{s}!${v}${a}:${v}${b})/'
                    'SUMPRODUCT(--({s}!$B${a}:$B${b}=$A{r}),'
                    '{s}!$D${a}:$D${b},--ISNUMBER({s}!${v}${a}:${v}${b})),"")'
                    .format(s=AUD, a=AU_INI, b=AU_FIN, r=r, v=vis), fmt=DEC2)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF(OR($D{r}="",$F{r}=""),"",$F{r}-$D{r}),"")'
                .format(r=r), fmt=DEC2)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF(OR($D{r}="",$F{r}=""),"",'
                'IF($F{r}-$D{r}>$D${t},"Mejora",'
                'IF($D{r}-$F{r}>$D${t},"Empeora","Igual"))),"")'
                .format(r=r, t=RA_TOL), bold=True)
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF(OR($F{r}="",$D${e}=""),"",$F{r}/$D${e}),"")'
                .format(r=r, e=RA_ESCALA), fmt=PCT)
        motor.verde(ws, 'J%d' % r)

    motor.val(ws, 'A%d' % RA_TOT, 'TOTAL (todas las áreas)', bold=True)
    motor.f(ws, 'B%d' % RA_TOT,
            '=COUNTIF({s}!$B${a}:$B${b},"<>")'.format(s=AUD, a=AU_INI,
                                                      b=AU_FIN), fmt=ENT,
            bold=True)
    motor.f(ws, 'C%d' % RA_TOT,
            '=IFERROR(SUM({s}!$D${a}:$D${b}),"")'.format(s=AUD, a=AU_INI,
                                                         b=AU_FIN), fmt=ENT,
            bold=True)
    for j, col in enumerate(COLS_RA):
        motor.f(ws, '%s%d' % (col, RA_TOT),
                '=IFERROR(IF({s}!{v}${p}="","",{s}!{v}${p}),"")'
                .format(s=AUD, v=COLS_VISITA[j], p=AU_POND), fmt=DEC2,
                bold=True)
    motor.f(ws, 'G%d' % RA_TOT,
            '=IFERROR(IF(OR($D{r}="",$F{r}=""),"",$F{r}-$D{r}),"")'
            .format(r=RA_TOT), fmt=DEC2, bold=True)
    motor.f(ws, 'H%d' % RA_TOT,
            '=IFERROR(IF(OR($D{r}="",$F{r}=""),"",'
            'IF($F{r}-$D{r}>$D${t},"Mejora",'
            'IF($D{r}-$F{r}>$D${t},"Empeora","Igual"))),"")'
            .format(r=RA_TOT, t=RA_TOL), bold=True)
    motor.f(ws, 'I%d' % RA_TOT,
            '=IFERROR(IF(OR($F{r}="",$D${e}=""),"",$F{r}/$D${e}),"")'
            .format(r=RA_TOT, e=RA_ESCALA), fmt=PCT, bold=True)
    for col in 'ABCDEFGHI':
        ws['%s%d' % (col, RA_TOT)].fill = PatternFill('solid', fgColor=CREMA)

    semaforo_3(ws, 'D%d:F%d' % (RA_INI, RA_TOT), '$D%d' % RA_INI,
               '$D$%d' % RA_UM_V, '$D$%d' % RA_UM_A)
    motor.semaforo_texto(ws, 'H%d:H%d' % (RA_INI, RA_TOT), (
        ('Mejora', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Empeora', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('Igual', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    motor.val(ws, 'A%d' % (RA_TOT + 2),
              'Mira la fila TOTAL y las seis de arriba a la vez. Si el total '
              'mejora y un área empeora, la mejora del resto está tapando un '
              'agujero que el cliente sí ve.', wrap=True)
    motor.val(ws, 'A%d' % (RA_TOT + 3),
              'La variación son PUNTOS de la escala, no porcentaje: la resta '
              'directa entre la tercera visita y la primera.', wrap=True)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Histórico»
# --------------------------------------------------------------------------
def hoja_historico(wb):
    ws = wb.create_sheet('Histórico')
    cabecera(ws, 'Histórico — una visita es una foto; tres son una tendencia')
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14
    for col in COLS_HI_AREA:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['L'].width = 46

    # Espejo de los dos umbrales: el formato condicional de Excel no puede
    # referirse a otra hoja, así que el semáforo de esta hoja necesita tener
    # los umbrales AQUÍ. Se leen de «Auditoría» a través de «Resumen por Área».
    motor.val(ws, 'C%d' % HI_UM_V, 'Umbral verde (0-5), en vigor')
    motor.f(ws, 'D%d' % HI_UM_V,
            '=IF({s}!$D${v}="","",{s}!$D${v})'.format(s=RES, v=RA_UM_V),
            fmt=DEC2)
    ws['D%d' % HI_UM_V].fill = PatternFill('solid', fgColor=GRIS_BG)
    motor.val(ws, 'C%d' % HI_UM_A, 'Umbral ámbar (0-5), en vigor')
    motor.f(ws, 'D%d' % HI_UM_A,
            '=IF({s}!$D${a}="","",{s}!$D${a})'.format(s=RES, a=RA_UM_A),
            fmt=DEC2)
    ws['D%d' % HI_UM_A].fill = PatternFill('solid', fgColor=GRIS_BG)
    motor.val(ws, 'F%d' % HI_UM_V,
              'Los umbrales se cambian en la hoja «Auditoría»; aquí sólo se '
              'leen.')

    encabezados(ws, HI_CAB, [
        ('A', 'Visita', None), ('B', 'Fecha', None), ('C', 'Auditor', None),
        ('D', 'Puntuación ponderada (0-5)', None),
        ('E', 'Cumplimiento (%)', None),
    ] + [(COLS_HI_AREA[i], DE.AREAS_AUDITORIA[i], None)
         for i in range(len(DE.AREAS_AUDITORIA))] + [('L', 'Notas', None)])

    for i in range(HI_FIN - HI_INI + 1):
        r = HI_INI + i
        vis = COLS_VISITA[i]
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.f(ws, 'B%d' % r,
                '=IF({s}!{v}${f}="","",{s}!{v}${f})'.format(s=AUD, v=vis,
                                                            f=AU_FECHA),
                fmt=FECHA)
        motor.f(ws, 'C%d' % r,
                '=IF({s}!{v}${a}="","",{s}!{v}${a})'.format(s=AUD, v=vis,
                                                            a=AU_AUDITOR))
        cel = motor.f(ws, 'D%d' % r,
                      '=IFERROR(IF({s}!{v}${p}="","",{s}!{v}${p}),"")'
                      .format(s=AUD, v=vis, p=AU_POND), fmt=DEC2, bold=True)
        cel.fill = PatternFill('solid', fgColor=CREMA)
        motor.f(ws, 'E%d' % r,
                '=IFERROR(IF({s}!{v}${c}="","",{s}!{v}${c}),"")'
                .format(s=AUD, v=vis, c=AU_PCT), fmt=PCT)
        for k, col in enumerate(COLS_HI_AREA):
            origen = COLS_RA[i]
            motor.f(ws, '%s%d' % (col, r),
                    '=IFERROR(IF({s}!{o}${f}="","",{s}!{o}${f}),"")'
                    .format(s=RES, o=origen, f=RA_INI + k), fmt=DEC2)
        motor.verde(ws, 'L%d' % r)

    semaforo_3(ws, 'D%d:D%d' % (HI_INI, HI_FIN), '$D%d' % HI_INI,
               '$D$%d' % HI_UM_V, '$D$%d' % HI_UM_A)
    semaforo_3(ws, '{a}{i}:{b}{f}'.format(a=COLS_HI_AREA[0], i=HI_INI,
                                          b=COLS_HI_AREA[-1], f=HI_FIN),
               '$F%d' % HI_INI, '$D$%d' % HI_UM_V, '$D$%d' % HI_UM_A)

    seccion(ws, 'A%d' % (HI_RES - 1), 'LECTURA — lo calcula el libro')
    motor.val(ws, 'B%d' % HI_RES, 'Visitas registradas')
    motor.f(ws, 'E%d' % HI_RES,
            '=COUNTIF($B${a}:$B${b},"<>")'.format(a=HI_INI, b=HI_FIN),
            fmt=ENT)
    motor.val(ws, 'B%d' % (HI_RES + 1),
              'Variación de la primera visita a la última (puntos)')
    motor.f(ws, 'E%d' % (HI_RES + 1),
            '=IFERROR(IF(OR($D${a}="",$D${b}=""),"",$D${b}-$D${a}),"")'
            .format(a=HI_INI, b=HI_FIN), fmt=DEC2)
    motor.val(ws, 'B%d' % (HI_RES + 2), 'Mejor área en la última visita')
    motor.f(ws, 'E%d' % (HI_RES + 2),
            '=IFERROR(INDEX({s}!$A${a}:$A${b},MATCH(MAX({s}!$F${a}:$F${b}),'
            '{s}!$F${a}:$F${b},0)),"")'.format(s=RES, a=RA_INI, b=RA_FIN))
    motor.val(ws, 'B%d' % (HI_RES + 3), 'Área más floja en la última visita')
    motor.f(ws, 'E%d' % (HI_RES + 3),
            '=IFERROR(INDEX({s}!$A${a}:$A${b},MATCH(MIN({s}!$F${a}:$F${b}),'
            '{s}!$F${a}:$F${b},0)),"")'.format(s=RES, a=RA_INI, b=RA_FIN))
    motor.val(ws, 'B%d' % (HI_RES + 4), 'Área que más mejora')
    motor.f(ws, 'E%d' % (HI_RES + 4),
            '=IFERROR(INDEX({s}!$A${a}:$A${b},MATCH(MAX({s}!$G${a}:$G${b}),'
            '{s}!$G${a}:$G${b},0)),"")'.format(s=RES, a=RA_INI, b=RA_FIN))
    motor.val(ws, 'B%d' % (HI_RES + 5), 'ÁREA QUE MÁS EMPEORA', bold=True)
    cel = motor.f(ws, 'E%d' % (HI_RES + 5),
                  '=IFERROR(INDEX({s}!$A${a}:$A${b},'
                  'MATCH(MIN({s}!$G${a}:$G${b}),{s}!$G${a}:$G${b},0)),"")'
                  .format(s=RES, a=RA_INI, b=RA_FIN), bold=True)
    cel.fill = PatternFill('solid', fgColor=CREMA)
    motor.val(ws, 'B%d' % (HI_RES + 6), 'Áreas que empeoran')
    motor.f(ws, 'E%d' % (HI_RES + 6),
            '=COUNTIF({s}!$H${a}:$H${b},"Empeora")'.format(s=RES, a=RA_INI,
                                                           b=RA_FIN),
            fmt=ENT)
    motor.regla_expresion(
        ws, 'E%d' % (HI_RES + 6),
        '=AND(ISNUMBER($E${r}),$E${r}>0)'.format(r=HI_RES + 6))

    motor.val(ws, 'B%d' % (HI_RES + 8),
              'El área que más empeora es la primera línea del plan de acción '
              'del trimestre siguiente. Con responsable y con fecha, no con '
              'buenas intenciones.', wrap=True)

    grafico = LineChart()
    grafico.title = 'Puntuación ponderada por área, visita a visita'
    grafico.style = 2
    grafico.height = 9.5
    grafico.width = 26
    datos = Reference(ws, min_col=6, max_col=11, min_row=HI_CAB,
                      max_row=HI_FIN)
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(Reference(ws, min_col=1, min_row=HI_INI,
                                     max_row=HI_FIN))
    grafico.y_axis.numFmt = DEC2
    grafico.y_axis.title = 'Puntuación ponderada (0-5)'
    grafico.x_axis.title = 'Visita'
    grafico.x_axis.delete = False
    grafico.y_axis.delete = False
    grafico.x_axis.tickLblPos = 'low'
    grafico.y_axis.majorGridlines = None
    ws.add_chart(grafico, 'A%d' % (HI_RES + 10))

    total = LineChart()
    total.title = 'Puntuación ponderada total del local'
    total.style = 2
    total.height = 7.5
    total.width = 26
    total.add_data(Reference(ws, min_col=4, max_col=4, min_row=HI_CAB,
                             max_row=HI_FIN), titles_from_data=True)
    total.set_categories(Reference(ws, min_col=1, min_row=HI_INI,
                                   max_row=HI_FIN))
    total.y_axis.numFmt = DEC2
    total.y_axis.title = 'Puntuación ponderada (0-5)'
    total.x_axis.title = 'Visita'
    total.x_axis.delete = False
    total.y_axis.delete = False
    total.x_axis.tickLblPos = 'low'
    total.y_axis.majorGridlines = None
    ws.add_chart(total, 'A%d' % (HI_RES + 30))

    setup(ws)
    return ws


# --------------------------------------------------------------------------
def mapa():
    celdas_au = {
        'Escala máxima de puntuación': 'D%d' % AU_ESCALA,
        'Umbral verde': 'D%d' % AU_UM_V,
        'Umbral ámbar': 'D%d' % AU_UM_A,
        'Tolerancia de la tendencia': 'D%d' % AU_TOL,
    }
    for j, col in enumerate(COLS_VISITA):
        n = j + 1
        celdas_au['Visita %d: fecha' % n] = '%s%d' % (col, AU_FECHA)
        celdas_au['Visita %d: auditor' % n] = '%s%d' % (col, AU_AUDITOR)
        celdas_au['Visita %d: puntos valorados' % n] = '%s%d' % (col, AU_VAL)
        celdas_au['Visita %d: peso valorado' % n] = '%s%d' % (col, AU_PESO)
        celdas_au['Visita %d: suma de peso por puntuación' % n] = \
            '%s%d' % (col, AU_SUMA)
        celdas_au['Visita %d: puntuación ponderada' % n] = \
            '%s%d' % (col, AU_POND)
        celdas_au['Visita %d: cumplimiento (%%)' % n] = '%s%d' % (col, AU_PCT)
        celdas_au['Visita %d: media simple' % n] = '%s%d' % (col, AU_SIMPLE)

    celdas_ra = {
        'Escala máxima (espejo)': 'D%d' % RA_ESCALA,
        'Umbral verde (espejo)': 'D%d' % RA_UM_V,
        'Umbral ámbar (espejo)': 'D%d' % RA_UM_A,
        'Tolerancia (espejo)': 'D%d' % RA_TOL,
        'TOTAL: puntos de control': 'B%d' % RA_TOT,
        'TOTAL: peso': 'C%d' % RA_TOT,
        'TOTAL: visita 1': 'D%d' % RA_TOT,
        'TOTAL: visita 2': 'E%d' % RA_TOT,
        'TOTAL: visita 3': 'F%d' % RA_TOT,
        'TOTAL: variación de la visita 1 a la 3': 'G%d' % RA_TOT,
        'TOTAL: tendencia': 'H%d' % RA_TOT,
        'TOTAL: cumplimiento visita 3 (%)': 'I%d' % RA_TOT,
    }
    for i, area in enumerate(DE.AREAS_AUDITORIA):
        r = RA_INI + i
        celdas_ra['%s: puntos de control' % area] = 'B%d' % r
        celdas_ra['%s: peso' % area] = 'C%d' % r
        celdas_ra['%s: visita 1' % area] = 'D%d' % r
        celdas_ra['%s: visita 2' % area] = 'E%d' % r
        celdas_ra['%s: visita 3' % area] = 'F%d' % r
        celdas_ra['%s: variación de la visita 1 a la 3' % area] = 'G%d' % r
        celdas_ra['%s: tendencia' % area] = 'H%d' % r
        celdas_ra['%s: cumplimiento visita 3 (%%)' % area] = 'I%d' % r

    celdas_hi = {
        'Umbral verde (espejo)': 'D%d' % HI_UM_V,
        'Umbral ámbar (espejo)': 'D%d' % HI_UM_A,
        'Visitas registradas': 'E%d' % HI_RES,
        'Variación de la primera visita a la última': 'E%d' % (HI_RES + 1),
        'Mejor área en la última visita': 'E%d' % (HI_RES + 2),
        'Área más floja en la última visita': 'E%d' % (HI_RES + 3),
        'Área que más mejora': 'E%d' % (HI_RES + 4),
        'Área que más empeora': 'E%d' % (HI_RES + 5),
        'Áreas que empeoran': 'E%d' % (HI_RES + 6),
    }
    for i in range(HI_FIN - HI_INI + 1):
        r = HI_INI + i
        celdas_hi['Visita %d: fecha' % (i + 1)] = 'B%d' % r
        celdas_hi['Visita %d: auditor' % (i + 1)] = 'C%d' % r
        celdas_hi['Visita %d: puntuación ponderada' % (i + 1)] = 'D%d' % r
        celdas_hi['Visita %d: cumplimiento (%%)' % (i + 1)] = 'E%d' % r

    return {
        'fichero': NOMBRE + '.xlsx',
        'producto': PRODUCTO,
        'hojas': {
            'Auditoría': {
                'celdas': celdas_au,
                'tablas': [
                    {'titulo': '60 puntos de control en 6 áreas, con peso y '
                               'tres visitas',
                     'cols': [['#', 'A', 'num'], ['Área', 'B', 'txt'],
                              ['Punto de control', 'C', 'txt'],
                              ['Peso (1-3)', 'D', 'num'],
                              ['Visita 1: puntuación (0-5)', 'E', 'num'],
                              ['Visita 1: observación', 'F', 'txt'],
                              ['Visita 2: puntuación (0-5)', 'G', 'num'],
                              ['Visita 2: observación', 'H', 'txt'],
                              ['Visita 3: puntuación (0-5)', 'I', 'num'],
                              ['Visita 3: observación', 'J', 'txt']],
                     'filas': [AU_INI, AU_FIN]},
                ],
            },
            'Resumen por Área': {
                'celdas': celdas_ra,
                'tablas': [
                    {'titulo': 'Puntuación ponderada por área y visita',
                     'cols': [['Área', 'A', 'txt'],
                              ['Puntos de control', 'B', 'num'],
                              ['Peso del área', 'C', 'num'],
                              ['Visita 1 (0-5)', 'D', 'num'],
                              ['Visita 2 (0-5)', 'E', 'num'],
                              ['Visita 3 (0-5)', 'F', 'num'],
                              ['Variación de la visita 1 a la 3', 'G', 'num'],
                              ['Tendencia', 'H', 'txt'],
                              ['Cumplimiento visita 3 (%)', 'I', 'pct1']],
                     'filas': [RA_INI, RA_TOT]},
                ],
            },
            'Histórico': {
                'celdas': celdas_hi,
                'tablas': [
                    {'titulo': 'Una fila por visita',
                     'cols': [['Visita', 'A', 'num'], ['Fecha', 'B', 'txt'],
                              ['Auditor', 'C', 'txt'],
                              ['Puntuación ponderada (0-5)', 'D', 'num'],
                              ['Cumplimiento (%)', 'E', 'pct1']]
                     + [[DE.AREAS_AUDITORIA[i], COLS_HI_AREA[i], 'num']
                        for i in range(len(DE.AREAS_AUDITORIA))],
                     'filas': [HI_INI, HI_FIN]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_auditoria(wb)
    hoja_resumen_area(wb)
    hoja_historico(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-22s %d' % (hoja, n))


if __name__ == '__main__':
    main()
