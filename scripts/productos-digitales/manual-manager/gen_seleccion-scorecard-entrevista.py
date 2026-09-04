#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_seleccion-scorecard-entrevista.py — Libro 4 del pack «Manual del Manager
de Restaurante» (SPEC §2.2, fila 4).

Genera `build/seleccion-scorecard-entrevista.xlsx`:

  Instrucciones · Scorecard · Comparativa de Candidatos ·
  Preguntas por Competencia

Sirve para comparar candidatos con el MISMO criterio en vez de con la
impresión del día, y para dejar por escrito por qué se contrató a alguien.

DECISIÓN IMPORTANTE DE DISEÑO: la media ponderada usa SOLO las competencias
VALORADAS. Una competencia que no se valoró se deja EN BLANCO, y su peso sale
del denominador; puntuar 0 es otra cosa muy distinta (es decir que la persona
no la tiene). El candidato B del ejemplo llega sin valorar los idiomas: su
media se calcula sobre 17 puntos de peso, no sobre 18.

DECISIONES TÉCNICAS
-------------------
* Cero constantes dentro de una fórmula: pesos, umbral y escala están en
  celdas.
* `SUMPRODUCT(pesos,puntuaciones)` y `SUMPRODUCT(pesos,--ISNUMBER(...))`, sin
  `IF` anidado dentro del SUMPRODUCT: es el idioma de la familia y el que
  entienden Excel, Sheets, Numbers y pycel. La celda vacía vale 0 en el
  producto y 0 en el peso, así que la media sale sola.
* `RANK` NO se usa: pycel no la implementa y el libro se quedaría sin valores
  cacheados. El puesto en el ranking se calcula contando cuántas medias son
  mayores que la propia, con `SUMPRODUCT` y guarda `ISNUMBER` (empates
  incluidos, igual que RANK).
* Funciones prohibidas (INDIRECT, COUNTA, PMT, OFFSET, XLOOKUP, LET, LAMBDA,
  matrices dinámicas): cero.
* El umbral se ESPEJA en «Comparativa de Candidatos» porque el formato
  condicional de Excel no puede referirse a otra hoja.
* Nota legal obligatoria en la hoja de preguntas: art. 9.5 de la Ley 15/2022.

Salida fija (sin argumentos):
`<carpeta>/build/seleccion-scorecard-entrevista.xlsx`
"""
import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/'
       'guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

PRODUCTO = 'manual-manager-restaurante'
motor.CTX['producto'] = PRODUCTO

# --------------------------------------------------------------------------
NOMBRE = 'seleccion-scorecard-entrevista'
TITULO_LIBRO = 'Selección: Scorecard y Entrevista Estructurada'
SUBTITULO = 'AI Chef Pro · aichef.pro — Manual del Manager de Restaurante'
SUBJECT = 'Manual del Manager de Restaurante · v1.0 · septiembre 2026'
VERSION = DE.VERSION_LINE
BIO = DE.BIO
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde: '
               'Revisar, Desproteger hoja (no tiene contraseña).')
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL, GRIS_BG = 'FFF8E1', '1565C0', 'F2F2F2'
PCT, ENT = motor.FMT_PCT, motor.FMT_ENT
DEC1 = '#,##0.0'
DEC2 = '#,##0.00'

VERIF = 'Verificado el 04-09-2026'
URL_LEY_15 = 'https://www.boe.es/buscar/act.php?id=BOE-A-2022-11589'

NOTA_LEGAL = (
    'LO QUE LA LEY NO TE DEJA PREGUNTAR. El art. 9.5 de la Ley 15/2022, de 12 '
    'de julio, integral para la igualdad de trato y la no discriminación, dice '
    'que el empleador no podrá preguntar sobre las condiciones de salud del '
    'aspirante al puesto. Ninguna de las 24 preguntas de esta hoja toca salud, '
    'embarazo, planes de maternidad o paternidad, cargas familiares, situación '
    'de pareja, religión, origen ni orientación sexual. No es una '
    'recomendación de estilo: es lo que separa una entrevista de una '
    'infracción. ' + VERIF + ' · Ley 15/2022, art. 9.5 · ' + URL_LEY_15)

NOTA_CONDUCTA = (
    'Todas las preguntas son de CONDUCTA («cuéntame una vez que...»), no de '
    'opinión. Se pregunta por lo que la persona ya ha hecho, que es lo que '
    'mejor predice lo que hará en tu servicio. Anota la respuesta en el '
    'momento: a la tercera entrevista del día ya no te acuerdas de quién dijo '
    'qué.')

# --- filas de la hoja «Scorecard» -----------------------------------------
SC_PUESTO = 4
SC_ALEH = 5
SC_JORNADA = 6
SC_UMBRAL = 7
SC_ESCALA = 8
SC_CAB = 11
SC_NOM = 12
SC_INI = 13
SC_FIN = SC_INI + len(DE.SELECCION['competencias']) - 1
SC_VAL = SC_FIN + 2
SC_PESO = SC_VAL + 1
SC_SUMA = SC_VAL + 2
SC_MEDIA = SC_VAL + 3
SC_SIMPLE = SC_VAL + 4
SC_REC = SC_VAL + 5
COLS_CAND = ['D', 'E', 'F', 'G']

# --- filas de la hoja «Comparativa de Candidatos» --------------------------
CO_UMBRAL = 4
CO_PUESTO = 5
CO_ESCALA = 6
CO_CAB = 8
CO_INI = 9
CO_FIN = CO_INI + len(COLS_CAND) - 1
CO_RES = CO_FIN + 2

# --- filas de la hoja «Preguntas por Competencia» --------------------------
PR_CAB = 8
PR_INI = 9
PR_FIN = PR_INI + len(DE.PREGUNTAS_COMPETENCIA) - 1
PR_RES = PR_FIN + 2

ESCALA = [
    (1, 'No lo ha hecho nunca o no lo demuestra'),
    (2, 'Lo ha hecho poco y con apoyo'),
    (3, 'Lo hace de forma autónoma en lo básico'),
    (4, 'Lo hace bien y resuelve situaciones difíciles'),
    (5, 'Lo hace muy bien y puede enseñárselo a otro'),
]


# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto, wrap=True)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9                      # A4
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezados(ws, fila, cols, alto=38):
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12, color=GOLD)


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. Antes de ver a nadie, escribe en «Scorecard» el puesto y las ocho '
    'competencias que de verdad importan para ESE puesto, con su peso de 1 a '
    '3. Definir el criterio después de conocer a los candidatos es la forma '
    'más fina de engañarse a uno mismo.',
    '2. Fija el umbral de recomendación (celda verde). Es la media ponderada '
    'por debajo de la cual no contratas, por muy simpática que haya sido la '
    'conversación.',
    '3. Pon el nombre de cada candidato en la fila de nombres y puntúa de 1 a '
    '5 en cada competencia, con la escala que tienes arriba a la vista.',
    '4. Si una competencia NO la valoraste, deja la celda VACÍA. Vacío es '
    '«no valorada» y su peso sale del cálculo. Un 0 significa otra cosa: que '
    'la persona no tiene esa competencia. No son lo mismo y el libro los trata '
    'distinto.',
    '5. La media ponderada, la media simple, el ranking y la recomendación se '
    'calculan solos y se ven juntos en «Comparativa de Candidatos».',
    '6. Usa «Preguntas por Competencia» durante la entrevista: tres preguntas '
    'por competencia, en el mismo orden para todo el mundo. Anota la respuesta '
    'en la columna verde y puntúa ahí mismo.',
    '7. Guarda el libro relleno. Si mañana alguien pregunta por qué se '
    'contrató a esa persona, la respuesta está escrita y con fecha.',
]

NOTAS = [
    'LO QUE NO SE PUEDE PREGUNTAR está en la primera pantalla de la hoja de '
    'preguntas: art. 9.5 de la Ley 15/2022. Léelo antes de la primera '
    'entrevista, no después de la primera denuncia.',
    'La entrevista estructurada no es un interrogatorio: es hacerle a todo el '
    'mundo las mismas preguntas para poder comparar. Lo que cambia entre '
    'candidatos son las respuestas, no las preguntas.',
    'El peso de cada competencia es TUYO. En un restaurante de carta con '
    'servicio en mesa pesa el aguante de un servicio de 75 cubiertos; en un '
    'local de barra pesará otra cosa. Cambia los pesos antes de puntuar, '
    'nunca después.',
    'Los cuatro candidatos sembrados son un EJEMPLO del proceso modelado del '
    'pack (camarero o camarera de sala). Bórralos y pon los tuyos.',
    'Dos candidatos pueden empatar. El libro lo dice sin maquillarlo: el '
    'ranking les da el mismo puesto, igual que haría una clasificación '
    'deportiva. Si tienes que elegir, mira la columna de observaciones y las '
    'competencias de peso 3.',
    'Cuando cierres el proceso, el paso siguiente es el onboarding: el Kit de '
    'Gestión de Personal trae la hoja de acogida de los primeros 30 días.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila), NOTA_LEGAL, wrap=True)
    fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Scorecard»
# --------------------------------------------------------------------------
def hoja_scorecard(wb):
    ws = wb.create_sheet('Scorecard')
    cabecera(ws, 'Scorecard — el mismo criterio para todos los candidatos')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 54
    ws.column_dimensions['C'].width = 10
    for col in COLS_CAND:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['H'].width = 46
    ws.column_dimensions['I'].width = 60

    seccion(ws, 'A3', 'EL PUESTO — celdas verdes')
    motor.val(ws, 'B%d' % SC_PUESTO, 'Puesto')
    motor.val(ws, 'D%d' % SC_PUESTO, DE.SELECCION['puesto'], verde_=True)
    motor.val(ws, 'B%d' % SC_ALEH, 'Encuadre en el ALEH VI')
    motor.val(ws, 'D%d' % SC_ALEH,
              '%s · %s' % (DE.SELECCION['area_aleh'],
                           DE.SELECCION['grupo_aleh']), verde_=True)
    motor.val(ws, 'I%d' % SC_ALEH,
              'El ALEH VI ordena la hostelería en SEIS áreas funcionales y '
              'TRES grupos profesionales. «Encargado», «director» o «gerente» '
              'son denominaciones de uso, no categorías del convenio: quien '
              'lleva el local se clasifica por las funciones que hace. '
              + VERIF + ' · ALEH VI (BOE-A-2023-6344, con la modificación '
              'BOE-A-2026-18630) · '
              'https://www.boe.es/buscar/doc.php?id=BOE-A-2023-6344',
              wrap=True)
    motor.val(ws, 'B%d' % SC_JORNADA, 'Jornada')
    motor.val(ws, 'D%d' % SC_JORNADA, DE.SELECCION['jornada'], verde_=True)
    motor.val(ws, 'B%d' % SC_UMBRAL,
              'Umbral de recomendación (media ponderada, 1-5)')
    motor.val(ws, 'D%d' % SC_UMBRAL, DE.SELECCION['umbral_recomendacion'],
              fmt=DEC2, verde_=True)
    motor.val(ws, 'I%d' % SC_UMBRAL,
              'Por debajo de esta media ponderada, el libro dice «No alcanza '
              'el umbral». Súbelo o bájalo según lo que puedas permitirte '
              'esperar, pero fíjalo ANTES de entrevistar.', wrap=True)
    motor.dv_numerica(ws, ['D%d' % SC_UMBRAL], minimo=1, maximo=5,
                      titulo='Umbral (1-5)',
                      mensaje='Escribe la media ponderada mínima para '
                              'recomendar a un candidato, entre 1 y 5.')
    motor.val(ws, 'B%d' % SC_ESCALA, 'Escala de puntuación')
    motor.val(ws, 'D%d' % SC_ESCALA,
              ' · '.join('%d = %s' % (n, t) for n, t in ESCALA), wrap=True)

    seccion(ws, 'A10', 'SCORECARD — puntúa de 1 a 5; deja VACÍO lo que no '
                       'valoraste')
    cand = DE.SELECCION['candidatos']
    encabezados(ws, SC_CAB, [
        ('A', '#', None), ('B', 'Competencia', None), ('C', 'Peso (1-3)', None),
        ('D', 'Candidato 1', None), ('E', 'Candidato 2', None),
        ('F', 'Candidato 3', None), ('G', 'Candidato 4', None),
        ('H', 'Observaciones', None), ('I', 'Notas', None),
    ])
    motor.val(ws, 'B%d' % SC_NOM, 'Nombre del candidato', bold=True)
    for j, col in enumerate(COLS_CAND):
        motor.val(ws, '%s%d' % (col, SC_NOM),
                  cand[j][0] if j < len(cand) else None, verde_=True,
                  bold=True)

    v_peso, v_punt = [], []
    for i, (comp, peso) in enumerate(DE.SELECCION['competencias']):
        r = SC_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, comp, wrap=True)
        motor.val(ws, 'C%d' % r, peso, fmt=ENT, verde_=True)
        v_peso.append('C%d' % r)
        for j, col in enumerate(COLS_CAND):
            punt = cand[j][1][i] if j < len(cand) else None
            if punt is not None:
                motor.val(ws, '%s%d' % (col, r), punt, fmt=ENT)
            else:
                ws['%s%d' % (col, r)].number_format = ENT
            v_punt.append('%s%d' % (col, r))
        motor.verde(ws, 'B%d:G%d' % (r, r))
        motor.verde(ws, 'H%d' % r)
    motor.dv_numerica(ws, v_peso, minimo=1, maximo=3, titulo='Peso (1-3)',
                      mensaje='1 conviene, 2 importa, 3 es imprescindible '
                              'para el puesto.')
    motor.dv_numerica(ws, v_punt, minimo=1, maximo=5, titulo='Puntuación (1-5)',
                      mensaje='Puntúa de 1 a 5 con la escala de arriba. Si no '
                              'valoraste esa competencia, deja la celda VACÍA: '
                              'su peso sale del cálculo.')

    seccion(ws, 'A%d' % (SC_VAL - 1), 'RESULTADO — lo calcula el libro')
    motor.val(ws, 'B%d' % SC_VAL, 'Competencias valoradas')
    motor.val(ws, 'B%d' % SC_PESO, 'Peso valorado')
    motor.val(ws, 'B%d' % SC_SUMA, 'Suma de peso por puntuación')
    motor.val(ws, 'B%d' % SC_MEDIA, 'MEDIA PONDERADA (1-5)', bold=True)
    motor.val(ws, 'B%d' % SC_SIMPLE, 'Media simple (1-5)')
    motor.val(ws, 'B%d' % SC_REC, 'Recomendación', bold=True)
    for col in COLS_CAND:
        motor.f(ws, '%s%d' % (col, SC_VAL),
                '=SUMPRODUCT(--ISNUMBER({c}${a}:{c}${b}))'
                .format(c=col, a=SC_INI, b=SC_FIN), fmt=ENT)
        motor.f(ws, '%s%d' % (col, SC_PESO),
                '=SUMPRODUCT($C${a}:$C${b},--ISNUMBER({c}${a}:{c}${b}))'
                .format(c=col, a=SC_INI, b=SC_FIN), fmt=ENT)
        motor.f(ws, '%s%d' % (col, SC_SUMA),
                '=IFERROR(SUMPRODUCT($C${a}:$C${b},{c}${a}:{c}${b}),"")'
                .format(c=col, a=SC_INI, b=SC_FIN), fmt=ENT)
        cel = motor.f(ws, '%s%d' % (col, SC_MEDIA),
                      '=IFERROR(IF({c}${p}=0,"",{c}${s}/{c}${p}),"")'
                      .format(c=col, p=SC_PESO, s=SC_SUMA), fmt=DEC2,
                      bold=True)
        cel.fill = PatternFill('solid', fgColor=CREMA)
        motor.f(ws, '%s%d' % (col, SC_SIMPLE),
                '=IFERROR(IF({c}${v}=0,"",AVERAGE({c}${a}:{c}${b})),"")'
                .format(c=col, v=SC_VAL, a=SC_INI, b=SC_FIN), fmt=DEC2)
        motor.f(ws, '%s%d' % (col, SC_REC),
                '=IFERROR(IF(OR({c}${m}="",$D${u}=""),"",'
                'IF({c}${m}>=$D${u},"Recomendado","No alcanza el umbral")),"")'
                .format(c=col, m=SC_MEDIA, u=SC_UMBRAL), bold=True)
    motor.semaforo_texto(ws, 'D%d:G%d' % (SC_REC, SC_REC), (
        ('Recomendado', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No alcanza el umbral', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    motor.semaforo_isnumber(ws, 'D%d:G%d' % (SC_MEDIA, SC_MEDIA),
                            'D$%d' % SC_MEDIA, '<', '$D$%d' % SC_UMBRAL)

    motor.val(ws, 'B%d' % (SC_REC + 2),
              'La media ponderada usa SOLO las competencias valoradas: mira '
              'la fila «Peso valorado». Si un candidato la tiene más baja que '
              'los demás, es que se le valoró menos, y compararlo con el resto '
              'es comparar dos cosas distintas.', wrap=True)
    motor.val(ws, 'B%d' % (SC_REC + 3),
              'La media simple está ahí para contrastar: cuando la ponderada y '
              'la simple se separan mucho, el candidato es fuerte justo en lo '
              'que más pesa, o al revés.', wrap=True)
    setup(ws, titulos='$%d:$%d' % (SC_CAB, SC_CAB))
    return ws


# --------------------------------------------------------------------------
# Hoja «Comparativa de Candidatos»
# --------------------------------------------------------------------------
def hoja_comparativa(wb):
    ws = wb.create_sheet('Comparativa de Candidatos')
    cabecera(ws, 'Comparativa — quién va primero y por cuánto')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 52

    seccion(ws, 'A3', 'PARÁMETROS EN VIGOR — se leen de la hoja «Scorecard»; '
                      'cámbialos allí, no aquí')
    motor.val(ws, 'B%d' % CO_UMBRAL,
              'Umbral de recomendación (media ponderada, 1-5)')
    motor.f(ws, 'D%d' % CO_UMBRAL,
            '=IF(Scorecard!$D${u}="","",Scorecard!$D${u})'.format(u=SC_UMBRAL),
            fmt=DEC2)
    ws['D%d' % CO_UMBRAL].fill = PatternFill('solid', fgColor=GRIS_BG)
    motor.val(ws, 'B%d' % CO_PUESTO, 'Puesto')
    motor.f(ws, 'D%d' % CO_PUESTO,
            '=IF(Scorecard!$D${p}="","",Scorecard!$D${p})'.format(p=SC_PUESTO))
    ws['D%d' % CO_PUESTO].fill = PatternFill('solid', fgColor=GRIS_BG)
    motor.val(ws, 'B%d' % CO_ESCALA, 'Competencias del scorecard')
    motor.f(ws, 'D%d' % CO_ESCALA,
            '=COUNTIF(Scorecard!$B${a}:$B${b},"<>")'.format(a=SC_INI,
                                                           b=SC_FIN), fmt=ENT)
    ws['D%d' % CO_ESCALA].fill = PatternFill('solid', fgColor=GRIS_BG)

    encabezados(ws, CO_CAB, [
        ('A', '#', None), ('B', 'Candidato', None),
        ('C', 'Media ponderada', None), ('D', 'Media simple', None),
        ('E', 'Competencias valoradas', None), ('F', 'Peso valorado', None),
        ('G', 'Puesto en el ranking', None), ('H', 'Recomendación', None),
        ('I', 'Notas', None),
    ])
    for i, col in enumerate(COLS_CAND):
        r = CO_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.f(ws, 'B%d' % r,
                '=IF(Scorecard!{c}${n}="","",Scorecard!{c}${n})'
                .format(c=col, n=SC_NOM))
        motor.f(ws, 'C%d' % r,
                '=IFERROR(IF(Scorecard!{c}${m}="","",Scorecard!{c}${m}),"")'
                .format(c=col, m=SC_MEDIA), fmt=DEC2, bold=True)
        motor.f(ws, 'D%d' % r,
                '=IFERROR(IF(Scorecard!{c}${m}="","",Scorecard!{c}${m}),"")'
                .format(c=col, m=SC_SIMPLE), fmt=DEC2)
        motor.f(ws, 'E%d' % r,
                '=IFERROR(Scorecard!{c}${v},"")'.format(c=col, v=SC_VAL),
                fmt=ENT)
        motor.f(ws, 'F%d' % r,
                '=IFERROR(Scorecard!{c}${p},"")'.format(c=col, p=SC_PESO),
                fmt=ENT)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF(NOT(ISNUMBER($C{r})),"",'
                'SUMPRODUCT(--ISNUMBER($C${a}:$C${b}),'
                '--($C${a}:$C${b}>$C{r}))+1),"")'
                .format(r=r, a=CO_INI, b=CO_FIN), fmt=ENT, bold=True)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF(OR($C{r}="",$D${u}=""),"",'
                'IF($C{r}>=$D${u},"Recomendado","No alcanza el umbral")),"")'
                .format(r=r, u=CO_UMBRAL), bold=True)
        motor.verde(ws, 'I%d' % r)
    motor.semaforo_texto(ws, 'H%d:H%d' % (CO_INI, CO_FIN), (
        ('Recomendado', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No alcanza el umbral', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    motor.semaforo_isnumber(ws, 'C%d:C%d' % (CO_INI, CO_FIN), '$C%d' % CO_INI,
                            '<', '$D$%d' % CO_UMBRAL)
    motor.regla_expresion(
        ws, 'G%d:G%d' % (CO_INI, CO_FIN),
        '=AND(ISNUMBER($G{a}),$G{a}=1)'.format(a=CO_INI),
        bg=motor.CF_VERDE_BG, fg=motor.CF_VERDE_FG)

    seccion(ws, 'A%d' % (CO_RES - 1), 'LECTURA — lo calcula el libro')
    motor.val(ws, 'B%d' % CO_RES, 'Candidatos evaluados')
    motor.f(ws, 'D%d' % CO_RES,
            '=COUNTIF($B${a}:$B${b},"<>")'.format(a=CO_INI, b=CO_FIN),
            fmt=ENT)
    motor.val(ws, 'B%d' % (CO_RES + 1), 'MEJOR CANDIDATO', bold=True)
    cel = motor.f(ws, 'D%d' % (CO_RES + 1),
                  '=IFERROR(INDEX($B${a}:$B${b},MATCH(MAX($C${a}:$C${b}),'
                  '$C${a}:$C${b},0)),"")'.format(a=CO_INI, b=CO_FIN),
                  bold=True)
    cel.fill = PatternFill('solid', fgColor=CREMA)
    motor.val(ws, 'B%d' % (CO_RES + 2), 'Mejor media ponderada')
    motor.f(ws, 'D%d' % (CO_RES + 2),
            '=IFERROR(MAX($C${a}:$C${b}),"")'.format(a=CO_INI, b=CO_FIN),
            fmt=DEC2)
    motor.val(ws, 'B%d' % (CO_RES + 3),
              'Distancia entre el primero y el segundo')
    motor.f(ws, 'D%d' % (CO_RES + 3),
            '=IFERROR(LARGE($C${a}:$C${b},1)-LARGE($C${a}:$C${b},2),"")'
            .format(a=CO_INI, b=CO_FIN), fmt=DEC2)
    motor.val(ws, 'B%d' % (CO_RES + 4), 'Candidatos que alcanzan el umbral')
    motor.f(ws, 'D%d' % (CO_RES + 4),
            '=COUNTIF($H${a}:$H${b},"Recomendado")'.format(a=CO_INI, b=CO_FIN),
            fmt=ENT)
    motor.val(ws, 'B%d' % (CO_RES + 5), 'Candidatos por debajo del umbral')
    motor.f(ws, 'D%d' % (CO_RES + 5),
            '=COUNTIF($H${a}:$H${b},"No alcanza el umbral")'
            .format(a=CO_INI, b=CO_FIN), fmt=ENT)

    motor.val(ws, 'B%d' % (CO_RES + 7),
              'Si la distancia entre el primero y el segundo es pequeña, el '
              'scorecard no ha decidido por ti: decide tú, y escribe en las '
              'notas por qué.', wrap=True)
    motor.val(ws, 'B%d' % (CO_RES + 8),
              'Si NADIE alcanza el umbral, la respuesta correcta es no '
              'contratar y volver a publicar la oferta. Contratar por debajo '
              'del criterio propio se paga en el servicio y en la rotación.',
              wrap=True)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Preguntas por Competencia»
# --------------------------------------------------------------------------
def hoja_preguntas(wb):
    ws = wb.create_sheet('Preguntas por Competencia')
    cabecera(ws, 'Preguntas por competencia — las mismas para todos')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 78
    ws.column_dimensions['D'].width = 52
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 20

    seccion(ws, 'A4', 'ANTES DE LA PRIMERA PREGUNTA')
    motor.val(ws, 'A5', NOTA_LEGAL, wrap=True)
    ws.row_dimensions[5].height = 58
    motor.val(ws, 'A6', NOTA_CONDUCTA, wrap=True)
    ws.row_dimensions[6].height = 30

    encabezados(ws, PR_CAB, [
        ('A', '#', None), ('B', 'Competencia', None), ('C', 'Pregunta', None),
        ('D', 'Respuesta y notas', None), ('E', 'Puntuación (1-5)', None),
        ('F', 'Candidato', None),
    ])
    ws.freeze_panes = 'C9'

    v_punt = []
    for i, (comp, preg) in enumerate(DE.PREGUNTAS_COMPETENCIA):
        r = PR_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, comp, wrap=True)
        motor.val(ws, 'C%d' % r, preg, wrap=True)
        ws['E%d' % r].number_format = ENT
        motor.verde(ws, 'D%d:E%d' % (r, r))
        v_punt.append('E%d' % r)
    motor.val(ws, 'F%d' % PR_INI, 'Escribe aquí a quién estás entrevistando y '
                                  'guarda una copia del libro por candidato.',
              wrap=True)
    motor.verde(ws, 'F%d' % PR_INI)
    motor.dv_numerica(ws, v_punt, minimo=1, maximo=5,
                      titulo='Puntuación (1-5)',
                      mensaje='Puntúa la respuesta de 1 a 5. Déjala vacía si '
                              'no hiciste esa pregunta.')

    seccion(ws, 'A%d' % (PR_RES - 1), 'RESUMEN DE LA ENTREVISTA — lo calcula '
                                      'el libro')
    motor.val(ws, 'B%d' % PR_RES, 'Preguntas del guion')
    motor.f(ws, 'E%d' % PR_RES,
            '=COUNTIF($C${a}:$C${b},"<>")'.format(a=PR_INI, b=PR_FIN),
            fmt=ENT)
    motor.val(ws, 'B%d' % (PR_RES + 1), 'Preguntas puntuadas')
    motor.f(ws, 'E%d' % (PR_RES + 1),
            '=SUMPRODUCT(--ISNUMBER($E${a}:$E${b}))'.format(a=PR_INI,
                                                            b=PR_FIN),
            fmt=ENT)
    motor.val(ws, 'B%d' % (PR_RES + 2),
              'Media de las respuestas puntuadas (1-5)', bold=True)
    cel = motor.f(ws, 'E%d' % (PR_RES + 2),
                  '=IFERROR(IF($E${n}=0,"",AVERAGE($E${a}:$E${b})),"")'
                  .format(n=PR_RES + 1, a=PR_INI, b=PR_FIN), fmt=DEC2,
                  bold=True)
    cel.fill = PatternFill('solid', fgColor=CREMA)
    motor.val(ws, 'B%d' % (PR_RES + 3), 'Cobertura del guion (%)')
    motor.f(ws, 'E%d' % (PR_RES + 3),
            '=IFERROR(IF($E${t}=0,"",$E${n}/$E${t}),"")'
            .format(t=PR_RES, n=PR_RES + 1), fmt=PCT)

    motor.val(ws, 'B%d' % (PR_RES + 5),
              'Esta media NO sustituye a la del scorecard: son 24 respuestas '
              'sin ponderar. El scorecard es el que decide, porque pesa cada '
              'competencia según lo que importa en tu casa.', wrap=True)
    motor.val(ws, 'B%d' % (PR_RES + 6),
              'Si una pregunta no te sirve, cámbiala. Lo que no se puede '
              'cambiar es hacer preguntas distintas a cada candidato: '
              'entonces no estás comparando, estás eligiendo por simpatía.',
              wrap=True)
    setup(ws, titulos='$%d:$%d' % (PR_CAB, PR_CAB))
    return ws


# --------------------------------------------------------------------------
def mapa():
    celdas_sc = {
        'Puesto': 'D%d' % SC_PUESTO,
        'Encuadre en el ALEH VI': 'D%d' % SC_ALEH,
        'Jornada': 'D%d' % SC_JORNADA,
        'Umbral de recomendación': 'D%d' % SC_UMBRAL,
    }
    nombres = [c[0] for c in DE.SELECCION['candidatos']]
    for j, col in enumerate(COLS_CAND):
        n = nombres[j] if j < len(nombres) else 'Candidato %d' % (j + 1)
        celdas_sc['%s: nombre' % n] = '%s%d' % (col, SC_NOM)
        celdas_sc['%s: competencias valoradas' % n] = '%s%d' % (col, SC_VAL)
        celdas_sc['%s: peso valorado' % n] = '%s%d' % (col, SC_PESO)
        celdas_sc['%s: suma de peso por puntuación' % n] = \
            '%s%d' % (col, SC_SUMA)
        celdas_sc['%s: media ponderada' % n] = '%s%d' % (col, SC_MEDIA)
        celdas_sc['%s: media simple' % n] = '%s%d' % (col, SC_SIMPLE)
        celdas_sc['%s: recomendación' % n] = '%s%d' % (col, SC_REC)
    for i, (comp, _p) in enumerate(DE.SELECCION['competencias']):
        celdas_sc['Peso de «%s»' % comp] = 'C%d' % (SC_INI + i)

    celdas_co = {
        'Umbral en vigor (espejo)': 'D%d' % CO_UMBRAL,
        'Puesto (espejo)': 'D%d' % CO_PUESTO,
        'Competencias del scorecard': 'D%d' % CO_ESCALA,
        'Candidatos evaluados': 'D%d' % CO_RES,
        'Mejor candidato': 'D%d' % (CO_RES + 1),
        'Mejor media ponderada': 'D%d' % (CO_RES + 2),
        'Distancia entre el primero y el segundo': 'D%d' % (CO_RES + 3),
        'Candidatos que alcanzan el umbral': 'D%d' % (CO_RES + 4),
        'Candidatos por debajo del umbral': 'D%d' % (CO_RES + 5),
    }
    for j in range(len(COLS_CAND)):
        n = nombres[j] if j < len(nombres) else 'Candidato %d' % (j + 1)
        celdas_co['%s: media ponderada' % n] = 'C%d' % (CO_INI + j)
        celdas_co['%s: puesto en el ranking' % n] = 'G%d' % (CO_INI + j)
        celdas_co['%s: recomendación' % n] = 'H%d' % (CO_INI + j)

    return {
        'fichero': NOMBRE + '.xlsx',
        'producto': PRODUCTO,
        'hojas': {
            'Scorecard': {
                'celdas': celdas_sc,
                'tablas': [
                    {'titulo': 'Scorecard: 8 competencias con peso y 4 '
                               'candidatos',
                     'cols': [['#', 'A', 'num'], ['Competencia', 'B', 'txt'],
                              ['Peso (1-3)', 'C', 'num'],
                              ['Candidato 1', 'D', 'num'],
                              ['Candidato 2', 'E', 'num'],
                              ['Candidato 3', 'F', 'num'],
                              ['Candidato 4', 'G', 'num'],
                              ['Observaciones', 'H', 'txt']],
                     'filas': [SC_INI, SC_FIN]},
                ],
            },
            'Comparativa de Candidatos': {
                'celdas': celdas_co,
                'tablas': [
                    {'titulo': 'Comparativa de los cuatro candidatos',
                     'cols': [['#', 'A', 'num'], ['Candidato', 'B', 'txt'],
                              ['Media ponderada', 'C', 'num'],
                              ['Media simple', 'D', 'num'],
                              ['Competencias valoradas', 'E', 'num'],
                              ['Peso valorado', 'F', 'num'],
                              ['Puesto en el ranking', 'G', 'num'],
                              ['Recomendación', 'H', 'txt']],
                     'filas': [CO_INI, CO_FIN]},
                ],
            },
            'Preguntas por Competencia': {
                'celdas': {
                    'Nota legal del art. 9.5 de la Ley 15/2022': 'A5',
                    'Preguntas del guion': 'E%d' % PR_RES,
                    'Preguntas puntuadas': 'E%d' % (PR_RES + 1),
                    # La media de las respuestas (E%d) nace VACÍA a propósito:
                    # la hoja de preguntas es un guion en blanco que se
                    # rellena durante la entrevista. No se cita como cifra.
                    'Cobertura del guion (%)': 'E%d' % (PR_RES + 3),
                },
                'tablas': [
                    {'titulo': '24 preguntas de conducta, tres por '
                               'competencia',
                     'cols': [['#', 'A', 'num'], ['Competencia', 'B', 'txt'],
                              ['Pregunta', 'C', 'txt'],
                              ['Respuesta y notas', 'D', 'txt'],
                              ['Puntuación (1-5)', 'E', 'num']],
                     'filas': [PR_INI, PR_FIN]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_scorecard(wb)
    hoja_comparativa(wb)
    hoja_preguntas(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-28s %d' % (hoja, n))


if __name__ == '__main__':
    main()
