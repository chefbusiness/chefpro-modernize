#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_calendario-cumplimiento-legal.py — Libro 5 del pack «Manual del Manager de
Restaurante» (SPEC `manual-manager-SPEC.md` §2.2, fila 5).

Genera `build/calendario-cumplimiento-legal.xlsx`:

  Instrucciones · Estado Normativo · Calendario y Vencimientos ·
  Documentación Obligatoria · Topes de Jornada · Permisos y Cómputo ·
  Régimen Disciplinario ALEH

Es el DIFERENCIADOR legal del producto (SPEC §2.4): de los 36 productos
censados del mercado sólo 2 tratan el cumplimiento legal español con
profundidad, y los dos lo hacen como trámite de apertura. Éste es la rutina.

DECISIONES TÉCNICAS
-------------------
* **Ni una cifra legal escrita de memoria.** El texto de las hojas de
  referencia sale de `auditorias/guias-v2-research-sector.json` por id `MM-*`
  (fuente_titulo + url + dato) y de `datos_ejemplo.py`. Cada fila que fija un
  dato legal lleva su columna «Fuente (norma)», su «URL» y una
  «Verificación» con el formato «Verificado el 04-09-2026 · norma · URL»
  (SPEC D12). Un `assert` aborta si el JSON no trae un id citado.
* **Cero constantes dentro de fórmulas.** Periodicidades, umbrales de aviso
  (30/60 días), jornada máxima y tope de horas extra viven en celdas verdes.
  Hasta el texto del semáforo se construye con `TEXT($C$6,"0")`: si el cliente
  cambia el umbral, la etiqueta cambia con él.
* **`TODAY()` en UNA sola celda de todo el libro** (`Calendario y
  Vencimientos!C5`), y detrás de una celda verde de simulación: dejar vacía la
  celda de simulación usa la fecha real; escribir una fecha permite ver el
  calendario «desde» otro día sin tocar ni un dato.
* **El semáforo NO se cuenta por su etiqueta.** `COUNTIF(rango,"< 30 días")`
  parece razonable y está ROTO: un criterio que empieza por «<» lo lee Excel
  (y pycel) como una COMPARACIÓN, no como un texto. Medido: devuelve 0 donde
  hay 3. Los contadores van sobre la columna NUMÉRICA de días con `COUNTIF` /
  `COUNTIFS`, y la etiqueta se queda para leerla.
* **Sin emojis.** El molde de origen (`kit-gestion-personal/07!Vencimientos`)
  pinta el semáforo con «❌🔴🟡🟢» dentro del texto de la celda. Aquí el color
  lo pone el formato condicional (con `ISNUMBER`, §1.6) y el texto es
  «VENCIDO / < 30 días / < 60 días / OK / No vence»: todo WinAnsi, nada que se
  rompa en un PDF ni en un visor sin fuente de emoji.
* Funciones prohibidas (`INDIRECT`, `COUNTA`, `PMT`, `OFFSET`, `XLOOKUP`,
  `LET`, `LAMBDA`): cero. `EDATE` sí — verificada contra pycel antes de usarla.
* «Sin dato» = `""`, nunca `0`; `IFERROR(...,"")` en todo cociente y en toda
  resta de fechas.

Salida fija (sin argumentos): `<carpeta>/build/calendario-cumplimiento-legal.xlsx`
"""
import json
import os
import sys
from datetime import date

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

motor.CTX['producto'] = 'manual-manager-restaurante'

# --------------------------------------------------------------------------
NOMBRE = 'calendario-cumplimiento-legal'
TITULO_LIBRO = 'Calendario de Cumplimiento Legal'
SUBTITULO = 'AI Chef Pro · aichef.pro — Manual del Manager de Restaurante'
SUBJECT = 'Manual del Manager de Restaurante · v1.0 · septiembre 2026'
VERSION = DE.VERSION_LINE
BIO = DE.BIO
DESPROTEGER = DE.NOTA_DESPROTEGER
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL = 'FFF8E1', '1565C0'
EUR, PCT, ENT = motor.FMT_EUR, motor.FMT_PCT, motor.FMT_ENT
FECHA = motor.FMT_FECHA

FECHA_CORTE = DE.RESTAURANTE['fecha_corte_normativa']            # '2026-09-04'
FECHA_CORTE_D = date(*map(int, FECHA_CORTE.split('-')))
FECHA_CORTE_TXT = FECHA_CORTE_D.strftime('%d-%m-%Y')             # '04-09-2026'

SI_NO = ['Sí', 'No']
DISPONIBLE = ['Sí', 'No', 'En trámite']

# --------------------------------------------------------------------------
# El JSON legal: única fuente del texto de norma y de la URL (SPEC §2.2/D12)
# --------------------------------------------------------------------------
RUTA_JSON = os.path.normpath(os.path.join(
    AQUI, '..', 'auditorias', 'guias-v2-research-sector.json'))
with open(RUTA_JSON, encoding='utf-8') as _fh:
    MM = {e['id']: e for e in json.load(_fh)['datos']
          if str(e.get('id', '')).startswith('MM-')}
assert len(MM) >= 55, 'El JSON no trae los ids MM-* esperados: %d' % len(MM)

# URLs de norma que el JSON ya publica. NO se teclea ninguna a mano: se leen de
# la entrada que cita esa misma norma, para que un cambio del JSON las arrastre.
URL_ET = MM['MM-01']['url']                 # Estatuto de los Trabajadores
URL_ALEH_MOD = MM['MM-13']['url']           # BOE-A-2026-18630 (04-09-2026)
URL_ALEH_BASE = MM['MM-11']['url']          # BOE-A-2023-6344 (ALEH VI)
assert 'boe.es' in URL_ET and 'boe.es' in URL_ALEH_MOD
FUENTE_ALEH_MOD = ('ALEH VI, modificación publicada por Resolución de la DGT '
                   'de 25-08-2026 (BOE-A-2026-18630)')


def fuente_de(mm_id, articulo=''):
    """(norma, url) de una fila legal. Con `mm_id` manda el JSON; sin él, se
    deriva de la FAMILIA de la norma citada en `datos_ejemplo.py`, y sólo hacia
    una URL que el propio JSON publica. Si no hay ninguna, «sin dato» = ''."""
    if mm_id:
        e = MM[mm_id]                        # KeyError adrede si el id no está
        return e['fuente_titulo'], e['url']
    art = articulo or ''
    if 'ALEH' in art:
        return FUENTE_ALEH_MOD, URL_ALEH_MOD
    if art.rstrip().endswith('ET') or ' ET' in art:
        return art + ' — Estatuto de los Trabajadores (RDLeg 2/2015)', URL_ET
    return '', ''


def verificacion(norma, url):
    """La nota del SPEC D12: «Verificado el 04-09-2026 · norma · URL»."""
    if not norma:
        return ''
    txt = 'Verificado el ' + FECHA_CORTE_TXT + ' · ' + norma
    if url:
        txt += ' · ' + url
    return txt


# --------------------------------------------------------------------------
# Ayudas de formato (mismas que el molde `guia-food-cost/gen_plan-accion…`)
# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto, wrap=True)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, landscape=True):
    ws.page_setup.paperSize = 9                     # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8


def encabezados(ws, fila, cols, alto=42):
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def bloque(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)


def total(ws, coord, contenido, fmt=None, formula=False):
    cel = (motor.f(ws, coord, contenido, fmt=fmt, bold=True) if formula
           else motor.val(ws, coord, contenido, fmt=fmt, bold=True))
    cel.fill = PatternFill('solid', fgColor=CREMA)
    return cel


def verde_propio(ws, fila, col_et, col_val, etiqueta, valor, fmt, nota,
                 col_nota=None):
    motor.val(ws, col_et + str(fila), etiqueta)
    motor.val(ws, col_val + str(fila), valor, fmt=fmt, verde_=True)
    if col_nota and nota:
        motor.val(ws, col_nota + str(fila), nota, wrap=True)
    return '$' + col_val + '$' + str(fila)


def bandas(ws, rango, reglas):
    """Varias reglas de formato condicional sobre EL MISMO rango.

    `motor.regla_expresion` limpia el rango antes de añadir, así que llamarla
    tres veces deja sólo la última. El libro se genera de cero en cada pasada,
    de modo que aquí no hay nada que limpiar: se apilan en orden con
    `stopIfTrue`.
    """
    for formula, bg, fg in reglas:
        ws.conditional_formatting.add(
            rango, FormulaRule(formula=[formula], stopIfTrue=True,
                               font=Font(color=fg, bold=True),
                               fill=PatternFill(start_color=bg, end_color=bg,
                                                fill_type='solid')))


def texto_filas(ws, fila_ini, fila_fin, alto=34):
    for r in range(fila_ini, fila_fin + 1):
        ws.row_dimensions[r].height = alto


def wrap(ws, rango):
    for fila in (ws[rango] if ':' in rango else [[ws[rango]]]):
        for cel in fila:
            cel.alignment = Alignment(vertical='top', wrap_text=True)


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. Empieza por «Estado Normativo»: siete normas que a día de hoy están EN '
    'MOVIMIENTO, con lo que obligan ahora mismo y lo que tienes que hacer tú. '
    'La fecha de corte de la revisión está en una celda verde: cuando la '
    'actualices, actualiza también la fila que haya cambiado.',
    '2. En «Calendario y Vencimientos» escribe la fecha de tu ÚLTIMA actuación '
    'de cada punto y la periodicidad en meses. El libro calcula la próxima '
    'fecha, los días que quedan y el semáforo. Los umbrales de aviso (rojo y '
    'ámbar) son dos celdas verdes: pon los tuyos.',
    '3. Mira la columna «¿Lo fija una norma estatal?» antes de contratar nada: '
    'sólo CUATRO familias de este calendario tienen la periodicidad fijada por '
    'una norma estatal. El resto es criterio de tu casa (o de tu seguro, o de '
    'tu plan de limpieza), y puedes cambiarlo.',
    '4. «Documentación Obligatoria» son los doce documentos que un inspector '
    'puede pedirte y dónde tienen que estar. Marca «Sí» a medida que los '
    'localizas: el porcentaje completado se calcula solo.',
    '5. «Topes de Jornada», «Permisos y Cómputo» y «Régimen Disciplinario '
    'ALEH» son hojas de CONSULTA: el dato, el artículo, la fuente y el enlace '
    'al BOE para que puedas comprobarlo tú. No se rellenan, se leen — salvo la '
    'pequeña calculadora de jornada y horas extra de la primera.',
    '6. Si quieres ver el calendario «desde» otro día (para preparar una '
    'inspección, o para saber qué vencerá en Navidad), escribe esa fecha en la '
    'celda verde de simulación de «Calendario y Vencimientos». Déjala vacía y '
    'el libro vuelve a usar la fecha de hoy.',
    '7. Lleva este libro a la reunión mensual con los vencimientos abiertos '
    'delante, y pasa lo que esté en rojo a la hoja «Actas y Acuerdos» del '
    'libro de reuniones: un vencimiento sin responsable y sin fecha no se '
    'resuelve.',
]

NOTAS = [
    'QUÉ FECHAS SON OBLIGACIÓN LEGAL Y CUÁLES NO. De los 18 puntos de control '
    'de este calendario, la periodicidad sólo la fija una norma estatal en '
    'CUATRO familias: el registro de jornada (que es diario), la inspección '
    'del ascensor, la revisión de la instalación de gas y los extintores (cuyos '
    'cuatro vencimientos son un mismo punto). Todo lo demás —control de '
    'plagas, limpieza de campana y conductos, verificación de termómetros, '
    'analítica de agua, formación de manipuladores y de prevención, revisión '
    'de la evaluación de riesgos, registro retributivo, plan de desperdicio, '
    'seguro— se vende muchas veces como «obligación legal cada X meses» y NO '
    'lo es: la ley te exige el RESULTADO (que el plan sea eficaz, que la '
    'medida sea fiable, que la formación esté acreditada), no el calendario. '
    'Las periodicidades sembradas ahí son criterio de la casa y son EDITABLES.',
    'ESTE LIBRO NO SUSTITUYE A UN ASESOR LABORAL NI A UN SERVICIO DE '
    'PREVENCIÓN. Es un calendario y un cuadro de consulta con la fuente '
    'primaria al lado para que puedas comprobar cada dato tú mismo. Antes de '
    'tomar una decisión que afecte a una persona (un despido, una sanción, la '
    'denegación de un permiso), consulta con tu asesoría.',
    'CADA FILA LEGAL TRAE SU FUENTE Y SU ENLACE. La columna «Verificación» '
    'dice «Verificado el ' + FECHA_CORTE_TXT + ' · norma · URL». Esa es la '
    'fecha en la que se leyó el texto en el BOE. Si compras el manual dentro '
    'de un año, comprueba en «Estado Normativo» qué ha cambiado: las '
    'actualizaciones del producto van incluidas.',
    'EL MARCO ES ESPAÑOL. Los topes de jornada, los permisos y el régimen '
    'disciplinario son los del Estatuto de los Trabajadores y del ALEH VI '
    '(acuerdo laboral estatal de hostelería). Si trabajas fuera de España, lo '
    'que te sirve es el MÉTODO y el calendario: todas las periodicidades, los '
    'umbrales y los valores son celdas editables.',
    'EL CONVENIO PROVINCIAL MANDA EN LAS TABLAS SALARIALES, PERO NO EN TODO. '
    'El ALEH VI fija la clasificación profesional, el periodo de prueba, los '
    'contratos formativos y el régimen disciplinario, y el convenio provincial '
    'NO los puede modificar. Las tablas salariales sí son provinciales: se '
    'consultan en el buscador REGCON del Ministerio de Trabajo.',
    'ESTE LIBRO NO CUBRE APPCC NI SEGURIDAD ALIMENTARIA. Los registros de '
    'temperaturas, la trazabilidad y los planes de prerrequisitos son otro '
    'producto (Pack APPCC): mezclarlos aquí daría un checklist que no sirve ni '
    'para una cosa ni para la otra. Lo que sí está aquí es CUÁNDO toca '
    'revisarlos y qué documentación tiene que estar en el local.',
    'LOS DATOS SEMBRADOS SON UN EJEMPLO MODELADO (el restaurante «La Encina», '
    'el mismo del resto del pack). Bórralos y pon los tuyos: las fechas de tu '
    'última actuación son lo único que este libro no puede saber.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, landscape=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Estado Normativo»
# --------------------------------------------------------------------------
EN0 = 8                                   # primera fila de datos
EN1 = EN0 + len(DE.ESTADO_NORMATIVO) + 3  # 7 filas + 4 libres

# Las dos normas que `datos_ejemplo.py` no cuelga de ningún id `MM-*` (no hay
# entrada en el JSON): se cita la norma que su propio texto nombra y la URL
# queda vacía = «sin dato». NUNCA se inventa un enlace del BOE.
NORMA_SIN_ID = {
    'Verifactu': 'RDL 15/2025, art. 3 (aplazamiento del Reglamento Verifactu)',
    'Prohibición de fumar en terrazas':
        'Ley 28/2005, art. 2.2 (definición de espacio al aire libre)',
}


def hoja_estado_normativo(wb):
    ws = wb.create_sheet('Estado Normativo')
    cabecera(ws, 'Estado normativo — las normas que se están moviendo')
    apunte(ws, 'C3',
           'Siete normas en movimiento a la fecha de corte. Lo que importa no '
           'es lo que se ha anunciado: es lo que OBLIGA hoy.')

    motor.val(ws, 'B5', 'Fecha de corte de la revisión normativa')
    motor.val(ws, 'C5', FECHA_CORTE_D, fmt=FECHA, verde_=True)
    motor.dv_fecha(ws, ['C5'])
    motor.val(ws, 'D5',
              'Es el día en que se leyó cada norma en el BOE. Cuando revises '
              'este cuadro, cambia esta fecha y la fila que haya cambiado.',
              wrap=True)

    encabezados(ws, EN0 - 1, [
        ('A', '#', 5), ('B', 'Norma', 34),
        ('C', 'Estado a la fecha de corte', 58),
        ('D', 'Qué obliga hoy y qué hace el manager', 58),
        ('E', 'Fecha de corte', 13), ('F', 'Fuente (norma)', 34),
        ('G', 'URL', 46), ('H', 'Ref.', 8), ('I', 'Verificación', 58),
    ])
    ws.freeze_panes = 'B' + str(EN0)

    for i in range(EN1 - EN0 + 1):
        r = EN0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.ESTADO_NORMATIVO):
            norma, estado, hace, _corte, mm_id = DE.ESTADO_NORMATIVO[i]
            fuente, url = fuente_de(mm_id)
            if not fuente:
                fuente = NORMA_SIN_ID.get(norma, '')
            motor.val(ws, 'B%d' % r, norma)
            motor.val(ws, 'C%d' % r, estado)
            motor.val(ws, 'D%d' % r, hace)
            motor.val(ws, 'F%d' % r, fuente)
            motor.val(ws, 'G%d' % r, url)
            motor.val(ws, 'H%d' % r, mm_id)
            motor.val(ws, 'I%d' % r, verificacion(fuente, url))
        motor.verde(ws, 'B%d:D%d' % (r, r))
        motor.f(ws, 'E%d' % r, '=IFERROR(IF($C$5="","",$C$5),"")', fmt=FECHA)
        wrap(ws, 'B%d:I%d' % (r, r))
    texto_filas(ws, EN0, EN1, alto=58)

    fr = EN1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    motor.val(ws, 'B%d' % (fr + 1), 'Normas en seguimiento')
    motor.f(ws, 'C%d' % (fr + 1),
            '=COUNTIF($B$%d:$B$%d,"<>")' % (EN0, EN1), fmt=ENT)
    motor.val(ws, 'B%d' % (fr + 2), 'Normas con id de investigación citado')
    motor.f(ws, 'C%d' % (fr + 2),
            '=COUNTIF($H$%d:$H$%d,"<>")' % (EN0, EN1), fmt=ENT)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Calendario y Vencimientos» — el corazón del libro
# --------------------------------------------------------------------------
CV0 = 10
CV1 = CV0 + len(DE.CUMPLIMIENTO) + 7      # 18 puntos + 8 filas libres
SIM, HOY, U_ROJO, U_AMBAR = '$C$4', '$C$5', '$C$6', '$C$7'


def hoja_calendario(wb):
    ws = wb.create_sheet('Calendario y Vencimientos')
    cabecera(ws, 'Calendario y vencimientos — qué toca, cuándo y quién lo manda')
    apunte(ws, 'F3',
           'Los siete primeros puntos son los de periodicidad fijada por norma '
           'estatal (cuatro familias). Del octavo en adelante, la periodicidad '
           'la eliges tú.')

    motor.val(ws, 'B4', 'Fecha para simular (déjala vacía y se usa la de hoy)')
    motor.val(ws, 'C4', None, fmt=FECHA, verde_=True)
    motor.dv_fecha(ws, ['C4'])
    motor.val(ws, 'D4',
              'Escribe aquí una fecha para ver el calendario «desde» ese día. '
              'Vacía = hoy.', wrap=True)

    motor.val(ws, 'B5', 'Hoy (fecha de referencia del calendario)', bold=True)
    motor.f(ws, 'C5', '=IFERROR(IF(%s="",TODAY(),%s),"")' % (SIM, SIM),
            fmt=FECHA, bold=True)

    motor.val(ws, 'B6', 'Aviso ROJO — faltan menos de (días)')
    motor.val(ws, 'C6', 30, fmt=ENT, verde_=True)
    motor.val(ws, 'B7', 'Aviso ÁMBAR — faltan menos de (días)')
    motor.val(ws, 'C7', 60, fmt=ENT, verde_=True)
    motor.dv_numerica(ws, ['C6', 'C7'], minimo=1, maximo=365,
                      titulo='Días de aviso',
                      mensaje='Escribe un número de días entre 1 y 365.')
    motor.val(ws, 'D6',
              'Los dos umbrales del semáforo. El texto de la columna «Estado» '
              'se escribe solo con el número que pongas aquí.', wrap=True)

    encabezados(ws, CV0 - 1, [
        ('A', '#', 5), ('B', 'Punto de control', 52),
        ('C', 'Última actuación', 15), ('D', 'Periodicidad (meses)', 12),
        ('E', '¿Lo fija una norma estatal?', 14),
        ('F', 'Próxima fecha', 14), ('G', 'Días hasta vencer', 12),
        ('H', 'Estado', 16), ('I', 'Fuente (norma)', 34), ('J', 'URL', 46),
        ('K', 'Ref.', 8), ('L', 'Verificación', 54), ('M', 'Nota', 56),
    ])
    ws.freeze_panes = 'B' + str(CV0)

    v_est, v_fecha, v_per = [], [], []
    for i in range(CV1 - CV0 + 1):
        r = CV0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.CUMPLIMIENTO):
            punto, _fam, ultima, per, estatal, mm_id, nota = DE.CUMPLIMIENTO[i]
            fuente, url = fuente_de(mm_id)
            motor.val(ws, 'B%d' % r, punto)
            motor.val(ws, 'C%d' % r,
                      date(*map(int, ultima.split('-'))), fmt=FECHA)
            motor.val(ws, 'D%d' % r, per, fmt=ENT)
            motor.val(ws, 'E%d' % r, estatal)
            motor.val(ws, 'I%d' % r, fuente)
            motor.val(ws, 'J%d' % r, url)
            motor.val(ws, 'K%d' % r, mm_id)
            motor.val(ws, 'L%d' % r, verificacion(fuente, url))
            motor.val(ws, 'M%d' % r, nota)
        else:
            ws['C%d' % r].number_format = FECHA
            ws['D%d' % r].number_format = ENT
        motor.verde(ws, 'B%d:E%d' % (r, r))
        motor.verde(ws, 'M%d' % r)
        v_est.append('E%d' % r)
        v_fecha.append('C%d' % r)
        v_per.append('D%d' % r)
        motor.f(ws, 'F%d' % r,
                '=IFERROR(IF(OR($C{r}="",$D{r}=""),"",'
                'IF($D{r}=0,"",EDATE($C{r},$D{r}))),"")'.format(r=r),
                fmt=FECHA)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF(OR($F{r}="",{hoy}=""),"",$F{r}-{hoy}),"")'
                .format(r=r, hoy=HOY), fmt=ENT)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF($B{r}="","",IF($D{r}=0,"No vence",'
                'IF($F{r}="","",IF($G{r}<0,"VENCIDO",'
                'IF($G{r}<={ur},"< "&TEXT({ur},"0")&" días",'
                'IF($G{r}<={ua},"< "&TEXT({ua},"0")&" días","OK")))))),"")'
                .format(r=r, ur=U_ROJO, ua=U_AMBAR))
        wrap(ws, 'B%d:M%d' % (r, r))
    texto_filas(ws, CV0, CV1, alto=32)

    motor.dv_lista(ws, v_est, SI_NO, titulo='Sí o No',
                   mensaje='¿La periodicidad de este punto la fija una norma '
                           'estatal? Elige Sí o No.')
    motor.dv_fecha(ws, v_fecha)
    motor.dv_numerica(ws, v_per, minimo=0, maximo=600,
                      titulo='Periodicidad (meses)',
                      mensaje='Escribe cada cuántos MESES se repite. Escribe 0 '
                              'si no es un vencimiento periódico (una '
                              'obligación diaria o un hito con fecha fija).')

    # Semáforo: el color lo pone el formato condicional sobre la columna
    # NUMÉRICA de días, con ISNUMBER (§1.6). El texto de «Estado» sólo se lee.
    rango_h = 'H%d:H%d' % (CV0, CV1)
    bandas(ws, rango_h, [
        ('=$H{r}="No vence"'.format(r=CV0), motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('=AND(ISNUMBER($G{r}),$G{r}<={ur})'.format(r=CV0, ur=U_ROJO),
         motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('=AND(ISNUMBER($G{r}),$G{r}<={ua})'.format(r=CV0, ua=U_AMBAR),
         motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('=ISNUMBER($G{r})'.format(r=CV0),
         motor.CF_VERDE_BG, motor.CF_VERDE_FG),
    ])
    motor.semaforo_isnumber(ws, 'G%d:G%d' % (CV0, CV1), '$G%d' % CV0,
                            operador='<', umbral='0')
    motor.semaforo_texto(ws, 'E%d:E%d' % (CV0, CV1), (
        ('Sí', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('No', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    # --- Resumen ----------------------------------------------------------
    fr = CV1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    B = '$B$%d:$B$%d' % (CV0, CV1)
    E = '$E$%d:$E$%d' % (CV0, CV1)
    G = '$G$%d:$G$%d' % (CV0, CV1)
    H = '$H$%d:$H$%d' % (CV0, CV1)
    filas = [
        ('Puntos de control registrados', '=COUNTIF(%s,"<>")' % B, ENT),
        ('Puntos con periodicidad fijada por una norma estatal',
         '=COUNTIF(%s,"Sí")' % E, ENT),
        ('Puntos cuya periodicidad es criterio de la casa',
         '=COUNTIF(%s,"No")' % E, ENT),
        ('Puntos sin vencimiento periódico (obligación diaria o hito con fecha)',
         '=COUNTIF(%s,"No vence")' % H, ENT),
    ]
    r = fr + 1
    for etiqueta, formula, fmt in filas:
        motor.val(ws, 'B%d' % r, etiqueta)
        motor.f(ws, 'D%d' % r, formula, fmt=fmt)
        r += 1
    c_reg, c_est, c_casa, c_novence = ('$D$%d' % (fr + 1), '$D$%d' % (fr + 2),
                                       '$D$%d' % (fr + 3), '$D$%d' % (fr + 4))
    motor.val(ws, 'B%d' % r, 'Puntos con vencimiento')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s="","",%s-%s),"")' % (c_reg, c_reg, c_novence),
            fmt=ENT)
    c_conv = '$D$%d' % r
    r += 1
    total(ws, 'B%d' % r, 'PUNTOS VENCIDOS')
    total(ws, 'D%d' % r, '=COUNTIF(%s,"<0")' % G, fmt=ENT, formula=True)
    c_venc = '$D$%d' % r
    motor.regla_expresion(ws, c_venc.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s>0)' % (c_venc, c_venc))
    r += 1
    motor.val(ws, 'B%d' % r, 'En aviso ROJO (vencen dentro del umbral rojo)')
    motor.f(ws, 'D%d' % r,
            '=COUNTIFS(%s,">="&0,%s,"<="&%s)' % (G, G, U_ROJO), fmt=ENT)
    c_rojo = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'En aviso ÁMBAR (entre los dos umbrales)')
    motor.f(ws, 'D%d' % r,
            '=COUNTIFS(%s,">"&%s,%s,"<="&%s)' % (G, U_ROJO, G, U_AMBAR),
            fmt=ENT)
    r += 1
    motor.val(ws, 'B%d' % r, 'En verde (OK)')
    motor.f(ws, 'D%d' % r, '=COUNTIFS(%s,">"&%s)' % (G, U_AMBAR), fmt=ENT)
    c_ok = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Puntos en verde (%)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_conv, c_ok, c_conv),
            fmt=PCT, bold=True)
    c_pct = '$D$%d' % r
    r += 2
    p_obj = verde_propio(
        ws, r, 'B', 'D', 'Objetivo de puntos en verde (%)', 0.90, PCT,
        'Por debajo de este porcentaje, la celda «Puntos en verde (%)» se pone '
        'en rojo. Es tu listón, no una obligación legal.', col_nota='F')
    motor.dv_porcentaje(ws, ['D%d' % r], titulo='Objetivo en verde',
                        prompt='Se escribe en tanto por uno: 0,90 = 90 %.')
    motor.regla_expresion(ws, c_pct.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<%s)' % (c_pct, c_pct, p_obj))
    r += 2
    apunte(ws, 'B%d' % r,
           'Recordatorio: que un punto salga en verde significa que la fecha '
           'que TÚ has escrito todavía no ha vencido, no que la periodicidad '
           'sea obligatoria. Mira siempre la columna «¿Lo fija una norma '
           'estatal?» y la nota de la fila.')
    ws.row_dimensions[r].height = 30
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Documentación Obligatoria»
# --------------------------------------------------------------------------
DO0 = 7
DO1 = DO0 + len(DE.DOCUMENTACION_OBLIGATORIA) + 3


def hoja_documentacion(wb):
    ws = wb.create_sheet('Documentación Obligatoria')
    cabecera(ws, 'Documentación obligatoria — qué tiene que estar y dónde')
    apunte(ws, 'C3',
           'Doce documentos que una inspección puede pedirte. Marca «Sí» '
           'cuando lo tengas LOCALIZADO, no cuando creas que existe.')

    encabezados(ws, DO0 - 1, [
        ('A', '#', 5), ('B', 'Documento', 58),
        ('C', 'Dónde debe estar', 50), ('D', 'Quién lo pide', 30),
        ('E', '¿Disponible en el local?', 15),
        ('F', 'Última comprobación', 15), ('G', 'Nota', 44),
    ])
    ws.freeze_panes = 'B' + str(DO0)

    v_disp, v_fecha = [], []
    for i in range(DO1 - DO0 + 1):
        r = DO0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.DOCUMENTACION_OBLIGATORIA):
            doc, donde, quien = DE.DOCUMENTACION_OBLIGATORIA[i]
            motor.val(ws, 'B%d' % r, doc)
            motor.val(ws, 'C%d' % r, donde)
            motor.val(ws, 'D%d' % r, quien)
        else:
            ws['F%d' % r].number_format = FECHA
        motor.verde(ws, 'B%d:G%d' % (r, r))
        ws['F%d' % r].number_format = FECHA
        v_disp.append('E%d' % r)
        v_fecha.append('F%d' % r)
        wrap(ws, 'B%d:G%d' % (r, r))
    texto_filas(ws, DO0, DO1, alto=32)
    motor.dv_lista(ws, v_disp, DISPONIBLE, titulo='Disponibilidad',
                   mensaje='Elige Sí, No o En trámite.')
    motor.dv_fecha(ws, v_fecha)
    motor.semaforo_texto(ws, 'E%d:E%d' % (DO0, DO1), (
        ('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('En trámite', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('No', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))

    B = '$B$%d:$B$%d' % (DO0, DO1)
    E = '$E$%d:$E$%d' % (DO0, DO1)
    fr = DO1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    r = fr + 1
    for etiqueta, formula in [
            ('Documentos registrados', '=COUNTIF(%s,"<>")' % B),
            ('Disponibles en el local («Sí»)', '=COUNTIF(%s,"Sí")' % E),
            ('En trámite', '=COUNTIF(%s,"En trámite")' % E),
            ('Que faltan («No»)', '=COUNTIF(%s,"No")' % E)]:
        motor.val(ws, 'B%d' % r, etiqueta)
        motor.f(ws, 'D%d' % r, formula, fmt=ENT)
        r += 1
    c_reg, c_si, c_tram, c_no = ['$D$%d' % (fr + k) for k in (1, 2, 3, 4)]
    motor.val(ws, 'B%d' % r, 'Sin contestar')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s="","",%s-%s-%s-%s),"")'
            % (c_reg, c_reg, c_si, c_tram, c_no), fmt=ENT)
    r += 1
    total(ws, 'B%d' % r, 'DOCUMENTACIÓN COMPLETADA (%)')
    total(ws, 'D%d' % r,
          '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_reg, c_si, c_reg),
          fmt=PCT, formula=True)
    c_pct = '$D$%d' % r
    motor.regla_expresion(ws, c_pct.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<1)' % (c_pct, c_pct))
    r += 2
    apunte(ws, 'B%d' % r,
           'La casilla se marca sola en «Sí» cuando lo has visto con tus ojos '
           'y sabes en qué carpeta está. «Creo que lo tiene la gestoría» es un '
           '«En trámite».')
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Topes de Jornada» (SPEC D4)
# --------------------------------------------------------------------------
TJ0 = 6
TJ1 = TJ0 + len(DE.TOPES_JORNADA) - 1

NOTAS_TOPES = {
    'Jornada máxima':
        'La reducción a 37,5 horas NO está vigente: el Congreso la rechazó el '
        '10 de septiembre de 2025 y devolvió el proyecto al Gobierno. Sigue el '
        'máximo de 40 horas semanales de promedio en cómputo anual.',
    'Jornada ordinaria diaria máxima':
        'Se puede superar sólo con distribución irregular pactada en convenio '
        'o por acuerdo con la representación de la plantilla, respetando '
        'siempre las 12 horas de descanso entre jornadas.',
    'Descanso mínimo entre jornadas':
        'Doce horas entre el fin de una jornada y el inicio de la siguiente. '
        'Es el tope que más se rompe en hostelería con el cierre de noche y la '
        'apertura de la mañana siguiente.',
    'Descanso semanal':
        'Día y medio ininterrumpido, acumulable en periodos de hasta 14 días.',
    'Descanso en jornada continuada de más de 6 h':
        'Quince minutos de descanso, que sólo cuentan como tiempo de trabajo '
        'efectivo si así lo establece el convenio o el contrato.',
    'Horas extraordinarias':
        'Ochenta horas al año por persona, reducidas en proporción en las '
        'jornadas parciales. No cuentan para el tope las compensadas con '
        'descanso dentro de los cuatro meses siguientes.',
    'Horas complementarias pactadas':
        'Sólo caben con pacto específico POR ESCRITO y en contratos de al '
        'menos 10 horas semanales.',
    'Horas complementarias voluntarias':
        'Un 15 % adicional, ampliable por convenio, que la persona trabajadora '
        'puede rechazar sin que ello sea causa de sanción.',
    'Registro de jornada':
        'Registro DIARIO con hora de inicio y de fin de cada persona. El real '
        'decreto del registro digital sigue en tramitación: hoy el papel y una '
        'hoja de cálculo son válidos.',
    'Conservación del registro':
        'Cuatro años, a disposición de la plantilla, de su representación y de '
        'la Inspección de Trabajo, en el propio centro de trabajo.',
    'Contrato a tiempo parcial sin registro':
        'Si no se lleva el registro diario del contrato a tiempo parcial, '
        'totalizado mensualmente y entregado con copia a la persona '
        'trabajadora, el contrato SE PRESUME celebrado a jornada completa.',
}


def hoja_topes(wb):
    ws = wb.create_sheet('Topes de Jornada')
    cabecera(ws, 'Topes de jornada — lo que la ley no te deja pasar')
    apunte(ws, 'C3',
           'Hoja de consulta: el dato, el artículo, la fuente y el enlace al '
           'BOE. Debajo hay una calculadora pequeña con tus propias cifras.')

    encabezados(ws, TJ0 - 1, [
        ('A', '#', 5), ('B', 'Concepto', 42), ('C', 'Valor', 52),
        ('D', 'Artículo', 20), ('E', 'Fuente (norma)', 34), ('F', 'URL', 46),
        ('G', 'Ref.', 8), ('H', 'Verificación', 56), ('I', 'Nota', 58),
    ])
    ws.freeze_panes = 'B' + str(TJ0)

    for i, (concepto, valor, mm_id, articulo) in enumerate(DE.TOPES_JORNADA):
        r = TJ0 + i
        fuente, url = fuente_de(mm_id, articulo)
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, concepto, bold=True)
        motor.val(ws, 'C%d' % r, valor)
        motor.val(ws, 'D%d' % r, articulo)
        motor.val(ws, 'E%d' % r, fuente)
        motor.val(ws, 'F%d' % r, url)
        motor.val(ws, 'G%d' % r, mm_id)
        motor.val(ws, 'H%d' % r, verificacion(fuente, url))
        motor.val(ws, 'I%d' % r, NOTAS_TOPES.get(concepto, ''))
        wrap(ws, 'B%d:I%d' % (r, r))
    texto_filas(ws, TJ0, TJ1, alto=44)

    # --- Calculadora ------------------------------------------------------
    fr = TJ1 + 2
    bloque(ws, 'A%d' % fr, 'CALCULADORA — con tus cifras, no con las nuestras')
    r = fr + 1
    p_max = verde_propio(
        ws, r, 'B', 'D',
        'Jornada máxima legal de promedio en cómputo anual (h/semana)', 40, ENT,
        'Art. 34.1 ET. Es una celda editable porque fuera de España el máximo '
        'es otro; en España, hoy, son 40.', col_nota='F')
    r += 1
    p_pact = verde_propio(
        ws, r, 'B', 'D', 'Horas semanales pactadas de promedio en tu casa', 40, ENT,
        'La jornada media que estás haciendo de verdad, no la del contrato.',
        col_nota='F')
    r += 1
    motor.val(ws, 'B%d' % r, 'Exceso sobre el máximo legal (h/semana)',
              bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(OR(%s="",%s=""),"",%s-%s),"")'
            % (p_pact, p_max, p_pact, p_max), fmt='#,##0.0', bold=True)
    c_exceso = '$D$%d' % r
    motor.regla_expresion(ws, c_exceso.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s>0)' % (c_exceso, c_exceso))
    r += 2
    p_tope = verde_propio(
        ws, r, 'B', 'D',
        'Tope legal de horas extraordinarias al año por persona (h)', 80, ENT,
        'Art. 35.2 ET. En las jornadas parciales se reduce en proporción: '
        'ajusta esta celda para esa persona.', col_nota='F')
    r += 1
    p_acum = verde_propio(
        ws, r, 'B', 'D',
        'Horas extraordinarias acumuladas este año por esa persona (h)', 0, ENT,
        'No cuentan las compensadas con descanso dentro de los cuatro meses '
        'siguientes a su realización.', col_nota='F')
    r += 1
    motor.val(ws, 'B%d' % r, 'Margen hasta el tope (h)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(OR(%s="",%s=""),"",%s-%s),"")'
            % (p_tope, p_acum, p_tope, p_acum), fmt='#,##0.0', bold=True)
    c_margen = '$D$%d' % r
    motor.regla_expresion(ws, c_margen.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<0)' % (c_margen, c_margen))
    r += 1
    motor.val(ws, 'B%d' % r, 'Tope consumido (%)')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(OR(%s="",%s=0),"",%s/%s),"")'
            % (p_acum, p_tope, p_acum, p_tope), fmt=PCT)
    c_cons = '$D$%d' % r
    r += 1
    p_aviso = verde_propio(
        ws, r, 'B', 'D', 'Avisar cuando se supere el (%) del tope', 0.80, PCT,
        'Por encima de este porcentaje la celda «Tope consumido (%)» se pone '
        'en rojo. Tu listón, no el de la ley.', col_nota='F')
    motor.dv_porcentaje(ws, ['D%d' % r], titulo='Aviso del tope',
                        prompt='Se escribe en tanto por uno: 0,80 = 80 %.')
    motor.regla_expresion(ws, c_cons.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s>=%s)' % (c_cons, c_cons,
                                                         p_aviso))
    motor.dv_numerica(ws, [p_max.replace('$', ''), p_pact.replace('$', '')],
                      minimo=0, maximo=80, titulo='Horas semanales',
                      mensaje='Escribe las horas semanales (0-80).')
    motor.dv_numerica(ws, [p_tope.replace('$', ''), p_acum.replace('$', '')],
                      minimo=0, maximo=2000, titulo='Horas al año',
                      mensaje='Escribe las horas al año (0-2000).')
    r += 2
    apunte(ws, 'B%d' % r,
           'Las horas extraordinarias son voluntarias salvo pacto o convenio, '
           'y en cualquier caso se pagan o se compensan con descanso. Lo que '
           'no existe es la hora que no se registra: sin registro, la '
           'Inspección presume la jornada completa.')
    ws.row_dimensions[r].height = 30
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Permisos y Cómputo» (SPEC D4 + D6)
# --------------------------------------------------------------------------
PC0 = 6
PC1 = PC0 + len(DE.PERMISOS) - 1

# Retribuido / quién lo pide / nota, fila a fila. El texto sale del `dato` de
# los ids MM-08, MM-26, MM-27 y MM-54 del JSON, no de memoria (SPEC §2.2).
DETALLE_PERMISOS = {
    'Vacaciones anuales': (
        'Sí',
        'La persona trabajadora; el calendario se fija de común acuerdo y se '
        'conoce con dos meses de antelación como mínimo',
        'Mínimo 30 días naturales al año, retribuidas y NO sustituibles por '
        'dinero salvo extinción del contrato.'),
    'Preaviso del calendario de vacaciones': (
        'No aplica',
        'La empresa: publica el calendario',
        'No es un permiso: es la obligación de la empresa de que el calendario '
        'se conozca con dos meses de antelación.'),
    'Fallecimiento de cónyuge, pareja de hecho o pariente hasta 2.º grado': (
        'Sí',
        'La persona trabajadora, avisando y justificándolo',
        'ERROR FRECUENTE: son 2 días, ampliables en 2 más si hay '
        'desplazamiento, NO 5. El RDL 5/2023 separó el fallecimiento del '
        'accidente o enfermedad grave, que sí son 5 días.'),
    'Accidente o enfermedad graves, hospitalización o intervención que precise reposo': (
        'Sí',
        'La persona trabajadora, avisando y justificándolo',
        'Cinco días. Alcanzan a la pareja de hecho y a cualquier conviviente '
        'que requiera cuidado efectivo.'),
    'Fuerza mayor familiar': (
        'Sí',
        'La persona trabajadora, avisando lo antes posible',
        'Hasta 4 días al año medidos en HORAS de ausencia, no en días '
        'completos.'),
    'Permiso parental hasta los 8 años': (
        'No',
        'La persona trabajadora, con el preaviso que fije el convenio',
        'Art. 48 bis ET: hasta 8 semanas, continuas o discontinuas, hasta que '
        'el menor cumple ocho años; derecho individual e intransferible y NO '
        'RETRIBUIDO, porque el art. 45.1.o) lo tipifica como causa de '
        'SUSPENSIÓN del contrato y el art. 45.2 exonera de las obligaciones '
        'recíprocas de trabajar y remunerar. Tampoco hay prestación de la '
        'Seguridad Social que lo cubra. NO es lo mismo que las dos semanas de '
        'la fila siguiente.'),
    'Permiso parental retribuido hasta los 8 años': (
        'Sí',
        'La persona trabajadora; la prestación se solicita a la Seguridad '
        'Social',
        'Art. 48.4.c) ET: dos semanas, cuatro en monoparentalidad. Desde el '
        'RDL 9/2025 forman parte del permiso por nacimiento. También llegan '
        'hasta los ocho años y también son derecho individual e '
        'intransferible, pero SÍ ESTÁN RETRIBUIDAS: el art. 177 LGSS declara '
        'situación protegida los descansos de los apartados 4, 5 y 6 del art. '
        '48 ET. Denegarlas o descontarlas confundiéndolas con el permiso '
        'parental del art. 48 bis es negar un derecho retribuido.'),
    'Nacimiento y cuidado de menor': (
        'Sí',
        'La persona trabajadora, con quince días de preaviso a la empresa',
        'Diecinueve semanas por progenitor desde el RDL 9/2025 (32 en '
        'monoparentalidad), no 16: 6 obligatorias e ininterrumpidas tras el '
        'parto, 11 distribuibles hasta los doce meses y 2 para el cuidado del '
        'menor hasta los ocho años. En adopción, guarda con fines de adopción '
        'y acogimiento, también 19 semanas por persona adoptante.'),
    'Guarda legal: reducción de jornada': (
        'No',
        'La persona trabajadora elige la concreción horaria dentro de su '
        'jornada ordinaria',
        'No es un permiso retribuido: es una reducción de jornada con '
        'reducción proporcional del salario. Quien elige el horario dentro de '
        'su jornada ordinaria es la persona trabajadora.'),
    'Adaptación de jornada por conciliación': (
        'No aplica',
        'La persona trabajadora, por escrito',
        'Si la empresa no contesta dentro del plazo legal, la adaptación se '
        'entiende CONCEDIDA. No confundir con la reducción por guarda legal: '
        'aquí no se reduce la jornada ni el salario, se cambia su '
        'distribución.'),
}


def hoja_permisos(wb):
    ws = wb.create_sheet('Permisos y Cómputo')
    cabecera(ws, 'Permisos y cómputo — qué se pide, cuánto dura y quién paga')
    apunte(ws, 'D3',
           'OJO con el permiso parental: son DOS figuras distintas que se '
           'solapan casi punto por punto y sólo una se paga. Van en dos filas '
           'seguidas a propósito.')

    encabezados(ws, PC0 - 1, [
        ('A', '#', 5), ('B', 'Permiso o figura', 44),
        ('C', 'Duración o cómputo', 50), ('D', '¿Retribuido?', 12),
        ('E', 'Quién lo pide y cómo', 46), ('F', 'Artículo', 26),
        ('G', 'Fuente (norma)', 40), ('H', 'URL', 46), ('I', 'Ref.', 8),
        ('J', 'Verificación', 56), ('K', 'Nota', 60),
    ])
    ws.freeze_panes = 'B' + str(PC0)

    for i, (permiso, duracion, mm_id, articulo) in enumerate(DE.PERMISOS):
        r = PC0 + i
        fuente, url = fuente_de(mm_id, articulo)
        retribuido, quien, nota = DETALLE_PERMISOS[permiso]
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, permiso, bold=True)
        motor.val(ws, 'C%d' % r, duracion)
        motor.val(ws, 'D%d' % r, retribuido)
        motor.val(ws, 'E%d' % r, quien)
        motor.val(ws, 'F%d' % r, articulo)
        motor.val(ws, 'G%d' % r, fuente)
        motor.val(ws, 'H%d' % r, url)
        motor.val(ws, 'I%d' % r, mm_id)
        motor.val(ws, 'J%d' % r, verificacion(fuente, url))
        motor.val(ws, 'K%d' % r, nota)
        wrap(ws, 'B%d:K%d' % (r, r))
    texto_filas(ws, PC0, PC1, alto=62)
    motor.semaforo_texto(ws, 'D%d:D%d' % (PC0, PC1), (
        ('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('No aplica', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    B = '$B$%d:$B$%d' % (PC0, PC1)
    D = '$D$%d:$D$%d' % (PC0, PC1)
    fr = PC1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    r = fr + 1
    for etiqueta, formula in [
            ('Figuras recogidas', '=COUNTIF(%s,"<>")' % B),
            ('Retribuidas', '=COUNTIF(%s,"Sí")' % D),
            ('No retribuidas', '=COUNTIF(%s,"No")' % D)]:
        motor.val(ws, 'B%d' % r, etiqueta)
        motor.f(ws, 'D%d' % r, formula, fmt=ENT)
        r += 1
    r += 1
    apunte(ws, 'B%d' % r,
           'Los días de permiso del Estatuto son MÍNIMOS: tu convenio '
           'provincial puede mejorarlos, nunca empeorarlos. Y la duración de '
           'los permisos por fallecimiento y por enfermedad grave se cuenta en '
           'días hábiles o naturales según lo diga el convenio: compruébalo '
           'antes de descontar un día.')
    ws.row_dimensions[r].height = 44
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Régimen Disciplinario ALEH» (SPEC D4 + D8)
# --------------------------------------------------------------------------
RD0 = 6
RD1 = RD0 + len(DE.REGIMEN_DISCIPLINARIO) - 1

SANCION = {
    'Leve': 'Amonestación verbal o escrita',
    'Grave': 'Suspensión de empleo y sueldo, según la escala de sanciones del '
             'ALEH VI',
    'Muy grave': 'Suspensión de empleo y sueldo o despido disciplinario, según '
                 'la escala de sanciones del ALEH VI',
}

NOTAS_REGIMEN = {
    'Falta LEVE: no registrar la jornada':
        'Novedad de la modificación del ALEH VI publicada el 04-09-2026: el '
        'incumplimiento del registro de jornada se tipifica por ESCALA. Dos '
        'incumplimientos en un mes son falta leve.',
    'Falta LEVE: uso no autorizado del móvil durante el servicio':
        'Novedad de la modificación del 04-09-2026. Para poder sancionar hace '
        'falta que la norma de uso del móvil esté comunicada por escrito a la '
        'plantilla.',
    'Falta GRAVE: no registrar la jornada':
        'Tres o cuatro incumplimientos en un mes. La escala sube desde la '
        'falta leve de dos.',
    'Falta GRAVE: fumar en zonas no permitidas':
        'Novedad de la modificación del 04-09-2026. Recuerda que una terraza '
        'es legalmente espacio al aire libre sólo si tiene como máximo dos '
        'paredes (Ley 28/2005, art. 2.2).',
    'Falta MUY GRAVE: no registrar la jornada':
        'Cinco o más incumplimientos en un mes. Es el tramo más alto de la '
        'escala del registro de jornada.',
    'Audiencia previa al despido disciplinario':
        'En vigor desde el 4 de septiembre de 2026 y hasta el 31 de diciembre '
        'de 2030. Antes de entregar la carta hay que informar a la persona de '
        'los hechos que se le imputan y de su posible calificación jurídica, y '
        'darle 2 días para contestar. La norma trae una EXCEPCIÓN («a menos '
        'que no pueda pedirse razonablemente a la empresa que le conceda esta '
        'posibilidad») y una REMISIÓN: la audiencia previa NO sustituye al '
        'resto de obligaciones del art. 55.1 ET para representantes y '
        'afiliados, se SUMA a ellas. Si se aparta del servicio a la persona '
        'durante esos días, son permiso retribuido. Incorpora la STS '
        '1250/2024, de 18 de noviembre, sobre el art. 7 del Convenio 158 OIT.',
    'Vigencia de la modificación del ALEH VI':
        'La modificación se publicó en el BOE número 219, de 4 de septiembre '
        'de 2026, y su vigencia acordada llega hasta el 31 de diciembre de '
        '2030. Antes de aplicar una sanción, comprueba que sigue vigente.',
    'Indemnización por despido objetivo':
        'Veinte días por año con tope de 12 mensualidades, puestos a '
        'disposición SIMULTÁNEAMENTE a la entrega de la carta y con preaviso '
        'de 15 días. No ponerla a disposición a la vez que la carta es el '
        'error que más despidos objetivos convierte en improcedentes.',
    'Indemnización por despido improcedente':
        'Treinta y tres días por año con tope de 24 mensualidades. Los '
        'contratos anteriores al 12 de febrero de 2012 llevan 45 días por el '
        'tramo previo más 33 por el posterior, con tope de 720 días.',
    'Periodo de prueba en hostelería':
        'Lo fija el ALEH estatal y el convenio provincial NO lo puede '
        'modificar. Indefinido: 90, 60 o 45 días naturales según grupo '
        'profesional; temporal de más de 3 meses: 75, 45 o 30; temporal de '
        'hasta 3 meses: 60, 30 o 15. Es NULO si la persona ya desempeñó las '
        'mismas funciones en la empresa.',
}

TIPO_FILA = {
    'Audiencia previa al despido disciplinario': 'Procedimiento',
    'Vigencia de la modificación del ALEH VI': 'Vigencia',
    'Indemnización por despido objetivo': 'Indemnización',
    'Indemnización por despido improcedente': 'Indemnización',
    'Periodo de prueba en hostelería': 'Contratación',
}


def _gravedad(etiqueta):
    if etiqueta.startswith('Falta MUY GRAVE'):
        return 'Muy grave'
    if etiqueta.startswith('Falta GRAVE'):
        return 'Grave'
    if etiqueta.startswith('Falta LEVE'):
        return 'Leve'
    return ''


def hoja_regimen(wb):
    ws = wb.create_sheet('Régimen Disciplinario ALEH')
    cabecera(ws, 'Régimen disciplinario del ALEH VI — qué es falta y qué no')
    apunte(ws, 'E3',
           'Modificación del ALEH VI publicada el 04-09-2026 (BOE-A-2026-18630), '
           'vigente hasta el 31-12-2030. Este cuadro recoge la TIPIFICACIÓN de '
           'las faltas, no la graduación de la sanción.')

    encabezados(ws, RD0 - 1, [
        ('A', '#', 5), ('B', 'Falta o figura', 44), ('C', 'Tipo', 15),
        ('D', 'Gravedad', 12), ('E', 'Umbral o detalle', 44),
        ('F', 'Sanción posible (escala del ALEH VI)', 42),
        ('G', 'Artículo', 32), ('H', 'Fuente (norma)', 42), ('I', 'URL', 46),
        ('J', 'Ref.', 8), ('K', 'Verificación', 56), ('L', 'Nota', 60),
    ])
    ws.freeze_panes = 'B' + str(RD0)

    for i, (falta, detalle, mm_id, articulo) in enumerate(
            DE.REGIMEN_DISCIPLINARIO):
        r = RD0 + i
        fuente, url = fuente_de(mm_id, articulo)
        grav = _gravedad(falta)
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, falta, bold=True)
        motor.val(ws, 'C%d' % r, TIPO_FILA.get(falta, 'Falta'))
        motor.val(ws, 'D%d' % r, grav)
        motor.val(ws, 'E%d' % r, detalle)
        motor.val(ws, 'F%d' % r, SANCION.get(grav, ''))
        motor.val(ws, 'G%d' % r, articulo)
        motor.val(ws, 'H%d' % r, fuente)
        motor.val(ws, 'I%d' % r, url)
        motor.val(ws, 'J%d' % r, mm_id)
        motor.val(ws, 'K%d' % r, verificacion(fuente, url))
        motor.val(ws, 'L%d' % r, NOTAS_REGIMEN.get(falta, ''))
        wrap(ws, 'B%d:L%d' % (r, r))
    texto_filas(ws, RD0, RD1, alto=62)
    motor.semaforo_texto(ws, 'D%d:D%d' % (RD0, RD1), (
        ('Muy grave', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('Grave', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Leve', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    B = '$B$%d:$B$%d' % (RD0, RD1)
    D = '$D$%d:$D$%d' % (RD0, RD1)
    fr = RD1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN Y VIGENCIA — lo calcula el libro')
    r = fr + 1
    for etiqueta, formula in [
            ('Filas del cuadro', '=COUNTIF(%s,"<>")' % B),
            ('Faltas leves tipificadas', '=COUNTIF(%s,"Leve")' % D),
            ('Faltas graves tipificadas', '=COUNTIF(%s,"Grave")' % D),
            ('Faltas muy graves tipificadas', '=COUNTIF(%s,"Muy grave")' % D)]:
        motor.val(ws, 'B%d' % r, etiqueta)
        motor.f(ws, 'E%d' % r, formula, fmt=ENT)
        r += 1
    r += 1
    p_desde = verde_propio(
        ws, r, 'B', 'E', 'Vigencia de la modificación del ALEH VI: desde',
        date(2026, 9, 4), FECHA,
        'BOE-A-2026-18630, BOE número 219 de 4 de septiembre de 2026.',
        col_nota='G')
    r += 1
    p_hasta = verde_propio(
        ws, r, 'B', 'E', 'Vigencia de la modificación del ALEH VI: hasta',
        date(2030, 12, 31), FECHA,
        'Vigencia acordada del ALEH VI. Pasada esa fecha, comprueba el texto '
        'nuevo antes de aplicar nada de este cuadro.', col_nota='G')
    motor.dv_fecha(ws, [p_desde.replace('$', ''), p_hasta.replace('$', '')])
    r += 1
    motor.val(ws, 'B%d' % r, 'Días de vigencia que quedan', bold=True)
    motor.f(ws, 'E%d' % r,
            "=IFERROR(IF(OR(%s=\"\",'Calendario y Vencimientos'!%s=\"\"),\"\","
            "%s-'Calendario y Vencimientos'!%s),\"\")"
            % (p_hasta, HOY, p_hasta, HOY), fmt=ENT, bold=True)
    c_dias = '$E$%d' % r
    motor.regla_expresion(ws, c_dias.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<0)' % (c_dias, c_dias))
    r += 2
    apunte(ws, 'B%d' % r,
           'La escala concreta de días de suspensión de empleo y sueldo la '
           'fija el capítulo de régimen disciplinario del ALEH VI y, en su '
           'caso, el convenio provincial. Antes de sancionar a nadie: '
           'audiencia previa, carta con los hechos y su calificación, y '
           'asesoría. Una sanción mal instruida se anula entera.')
    ws.row_dimensions[r].height = 44
    setup(ws)
    return ws


# --------------------------------------------------------------------------
def mapa():
    cvr = CV1 + 3     # primera fila del resumen del calendario
    dor = DO1 + 3
    return {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Estado Normativo': {
                'celdas': {
                    'Fecha de corte de la revisión normativa': 'C5',
                    'Normas en seguimiento': 'C%d' % (EN1 + 3),
                    'Normas con id de investigación citado': 'C%d' % (EN1 + 4),
                },
                'tablas': [
                    {'titulo': 'Las 7 normas en movimiento a 04-09-2026',
                     'cols': [['#', 'A', 'num'], ['Norma', 'B', 'txt'],
                              ['Estado a la fecha de corte', 'C', 'txt'],
                              ['Qué obliga hoy y qué hace el manager', 'D',
                               'txt'],
                              ['Fecha de corte', 'E', 'txt'],
                              ['Fuente (norma)', 'F', 'txt'],
                              ['URL', 'G', 'txt'], ['Ref.', 'H', 'txt'],
                              ['Verificación', 'I', 'txt']],
                     'filas': [EN0, EN1]},
                ],
            },
            'Calendario y Vencimientos': {
                'celdas': {
                    'Hoy (fecha de referencia del calendario)': 'C5',
                    'Aviso ROJO — faltan menos de (días)': 'C6',
                    'Aviso ÁMBAR — faltan menos de (días)': 'C7',
                    'Puntos de control registrados': 'D%d' % cvr,
                    'Puntos con periodicidad fijada por una norma estatal':
                        'D%d' % (cvr + 1),
                    'Puntos cuya periodicidad es criterio de la casa':
                        'D%d' % (cvr + 2),
                    'Puntos sin vencimiento periódico': 'D%d' % (cvr + 3),
                    'Puntos con vencimiento': 'D%d' % (cvr + 4),
                    'PUNTOS VENCIDOS': 'D%d' % (cvr + 5),
                    'En aviso ROJO': 'D%d' % (cvr + 6),
                    'En aviso ÁMBAR': 'D%d' % (cvr + 7),
                    'En verde (OK)': 'D%d' % (cvr + 8),
                    'Puntos en verde (%)': 'D%d' % (cvr + 9),
                    'Objetivo de puntos en verde (%)': 'D%d' % (cvr + 11),
                },
                'tablas': [
                    {'titulo': '18 puntos de control y 8 filas libres',
                     'cols': [['#', 'A', 'num'],
                              ['Punto de control', 'B', 'txt'],
                              ['Última actuación', 'C', 'txt'],
                              ['Periodicidad (meses)', 'D', 'num'],
                              ['¿Lo fija una norma estatal?', 'E', 'txt'],
                              ['Próxima fecha', 'F', 'txt'],
                              ['Días hasta vencer', 'G', 'num'],
                              ['Estado', 'H', 'txt'],
                              ['Fuente (norma)', 'I', 'txt'],
                              ['URL', 'J', 'txt'], ['Ref.', 'K', 'txt'],
                              ['Verificación', 'L', 'txt'],
                              ['Nota', 'M', 'txt']],
                     'filas': [CV0, CV1]},
                ],
            },
            'Documentación Obligatoria': {
                'celdas': {
                    'Documentos registrados': 'D%d' % dor,
                    'Disponibles en el local': 'D%d' % (dor + 1),
                    'En trámite': 'D%d' % (dor + 2),
                    'Que faltan': 'D%d' % (dor + 3),
                    'Sin contestar': 'D%d' % (dor + 4),
                    'DOCUMENTACIÓN COMPLETADA (%)': 'D%d' % (dor + 5),
                },
                'tablas': [
                    {'titulo': '12 documentos obligatorios y 4 filas libres',
                     'cols': [['#', 'A', 'num'], ['Documento', 'B', 'txt'],
                              ['Dónde debe estar', 'C', 'txt'],
                              ['Quién lo pide', 'D', 'txt'],
                              ['¿Disponible en el local?', 'E', 'txt'],
                              ['Última comprobación', 'F', 'txt'],
                              ['Nota', 'G', 'txt']],
                     'filas': [DO0, DO1]},
                ],
            },
            'Topes de Jornada': {
                'celdas': {
                    'Jornada máxima legal de promedio (h/semana)':
                        'D%d' % (TJ1 + 3),
                    'Horas semanales pactadas de promedio': 'D%d' % (TJ1 + 4),
                    'Exceso sobre el máximo legal (h/semana)':
                        'D%d' % (TJ1 + 5),
                    'Tope legal de horas extraordinarias al año (h)':
                        'D%d' % (TJ1 + 7),
                    'Horas extraordinarias acumuladas (h)': 'D%d' % (TJ1 + 8),
                    'Margen hasta el tope (h)': 'D%d' % (TJ1 + 9),
                    'Tope consumido (%)': 'D%d' % (TJ1 + 10),
                    'Avisar cuando se supere el (%) del tope':
                        'D%d' % (TJ1 + 11),
                },
                'tablas': [
                    {'titulo': 'Los 11 topes de jornada, con norma y URL',
                     'cols': [['#', 'A', 'num'], ['Concepto', 'B', 'txt'],
                              ['Valor', 'C', 'txt'], ['Artículo', 'D', 'txt'],
                              ['Fuente (norma)', 'E', 'txt'],
                              ['URL', 'F', 'txt'], ['Ref.', 'G', 'txt'],
                              ['Verificación', 'H', 'txt'],
                              ['Nota', 'I', 'txt']],
                     'filas': [TJ0, TJ1]},
                ],
            },
            'Permisos y Cómputo': {
                'celdas': {
                    'Figuras recogidas': 'D%d' % (PC1 + 3),
                    'Retribuidas': 'D%d' % (PC1 + 4),
                    'No retribuidas': 'D%d' % (PC1 + 5),
                },
                'tablas': [
                    {'titulo': 'Permisos y cómputo, con las DOS figuras del '
                               'permiso parental',
                     'cols': [['#', 'A', 'num'],
                              ['Permiso o figura', 'B', 'txt'],
                              ['Duración o cómputo', 'C', 'txt'],
                              ['¿Retribuido?', 'D', 'txt'],
                              ['Quién lo pide y cómo', 'E', 'txt'],
                              ['Artículo', 'F', 'txt'],
                              ['Fuente (norma)', 'G', 'txt'],
                              ['URL', 'H', 'txt'], ['Ref.', 'I', 'txt'],
                              ['Verificación', 'J', 'txt'],
                              ['Nota', 'K', 'txt']],
                     'filas': [PC0, PC1]},
                ],
            },
            'Régimen Disciplinario ALEH': {
                'celdas': {
                    'Filas del cuadro': 'E%d' % (RD1 + 3),
                    'Faltas leves tipificadas': 'E%d' % (RD1 + 4),
                    'Faltas graves tipificadas': 'E%d' % (RD1 + 5),
                    'Faltas muy graves tipificadas': 'E%d' % (RD1 + 6),
                    'Vigencia del ALEH VI: desde': 'E%d' % (RD1 + 8),
                    'Vigencia del ALEH VI: hasta': 'E%d' % (RD1 + 9),
                    'Días de vigencia que quedan': 'E%d' % (RD1 + 10),
                },
                'tablas': [
                    {'titulo': 'Régimen disciplinario del ALEH VI tras la '
                               'modificación del 04-09-2026',
                     'cols': [['#', 'A', 'num'],
                              ['Falta o figura', 'B', 'txt'],
                              ['Tipo', 'C', 'txt'], ['Gravedad', 'D', 'txt'],
                              ['Umbral o detalle', 'E', 'txt'],
                              ['Sanción posible', 'F', 'txt'],
                              ['Artículo', 'G', 'txt'],
                              ['Fuente (norma)', 'H', 'txt'],
                              ['URL', 'I', 'txt'], ['Ref.', 'J', 'txt'],
                              ['Verificación', 'K', 'txt'],
                              ['Nota', 'L', 'txt']],
                     'filas': [RD0, RD1]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_estado_normativo(wb)
    hoja_calendario(wb)
    hoja_documentacion(wb)
    hoja_topes(wb)
    hoja_permisos(wb)
    hoja_regimen(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-30s %d' % (hoja, n))


if __name__ == '__main__':
    main()
