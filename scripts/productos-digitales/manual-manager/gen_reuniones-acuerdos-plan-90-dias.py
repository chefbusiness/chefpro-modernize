#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_reuniones-acuerdos-plan-90-dias.py — Libro 6 del pack «Manual del Manager
de Restaurante» (SPEC `manual-manager-SPEC.md` §2.2, fila 6).

Genera `build/reuniones-acuerdos-plan-90-dias.xlsx`:

  Instrucciones · Calendario de Reuniones · Guion de Reunión Semanal ·
  Uno-a-uno · Actas y Acuerdos · Plan 90 Días

Lo que hace único a este libro no es el acta: es que la reunión salga con
DECISIONES QUE TIENEN FECHA, y que esas decisiones se puedan ver vencidas de un
vistazo. La hoja «Plan 90 Días» recoge las que sobreviven al trimestre.

DECISIONES DE LA SPEC QUE SE APLICAN AQUÍ
-----------------------------------------
* **Sin hoja de briefing** (D3). El briefing de servicio es DIARIO y ya existe
  veinte veces en el catálogo (`kit-tareas/BONUS-01`,
  `kit-gestion-personal/BONUS-01`). Este libro cubre las tres cadencias que no
  cubre nadie: semanal, mensual y uno-a-uno.
* **La hoja «Plan 90 Días» NO es el plan de 90 días de la Guía Food Cost**
  (D3). Aquel decide qué pasa con cada PLATO de la carta; éste ordena las
  salidas de los siete libros del manual por área, semana y responsable. Se
  usan EN PARALELO. Lo dice la hoja «Instrucciones» en voz alta.
* Sin hoja de KPI: viven en `cuadro-de-mando-semanal-manager.xlsx` (libro 1).

DECISIONES TÉCNICAS
-------------------
* **`dv_lista` NO sirve para la lista de herramientas del pack.** Construye
  `formula1='"' + ','.join(opciones) + '"'` y dos de los siete nombres llevan
  coma dentro («Quejas, reclamaciones y reseñas», «Reuniones, acuerdos y plan
  de 90 días»): el desplegable saldría partido en once opciones sin avisar de
  nada. Esas dos listas se montan sobre un RANGO de la propia hoja.
* **`TODAY()` en UNA sola celda** (`Actas y Acuerdos!C5`), detrás de una celda
  verde de simulación; el resto del libro la referencia, incluido «Plan 90
  Días» por referencia entre hojas.
* Cero constantes dentro de fórmulas: la ventana de «vence esta semana», la
  fecha de inicio del plan y el número de semanas son celdas verdes. El `7` de
  «siete días por semana» sí es constante: es la definición de semana, no un
  parámetro del negocio.
* «Cerrado en plazo» necesita una fecha de cierre REAL, que el juego de datos
  no tiene: la columna existe, está vacía y su porcentaje devuelve `""` («sin
  dato»), nunca un 0 que se leería como «ninguno se cerró en plazo».
* Funciones prohibidas (`INDIRECT`, `COUNTA`, `PMT`, `OFFSET`, `XLOOKUP`,
  `LET`, `LAMBDA`): cero.

Salida fija (sin argumentos):
`<carpeta>/build/reuniones-acuerdos-plan-90-dias.xlsx`
"""
import json
import os
import sys
from datetime import date

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

motor.CTX['producto'] = 'manual-manager-restaurante'

# --------------------------------------------------------------------------
NOMBRE = 'reuniones-acuerdos-plan-90-dias'
TITULO_LIBRO = 'Reuniones, Acuerdos y Plan de 90 Días'
SUBTITULO = 'AI Chef Pro · aichef.pro — Manual del Manager de Restaurante'
SUBJECT = 'Manual del Manager de Restaurante · v1.0 · septiembre 2026'
VERSION = DE.VERSION_LINE
BIO = DE.BIO
DESPROTEGER = DE.NOTA_DESPROTEGER
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL = 'FFF8E1', '1565C0'
EUR, PCT, ENT = motor.FMT_EUR, motor.FMT_PCT, motor.FMT_ENT
FECHA = motor.FMT_FECHA

# Personas: «P01 · Marta L.». Ningún nombre lleva coma, así que esta lista sí
# puede ir en un desplegable de `dv_lista`.
PERSONAS = ['%s · %s' % (p[0], p[1]) for p in DE.PLANTILLA]
POR_ID = {p[0]: '%s · %s' % (p[0], p[1]) for p in DE.PLANTILLA}
assert not any(',' in x for x in PERSONAS), 'Un nombre con coma rompe la DV'

TIPOS_REUNION = ['Semanal de equipo', 'Mensual de resultados', 'Uno-a-uno']
CADENCIAS = ['Semanal', 'Mensual']
ESTADOS_REUNION = ['Prevista', 'Celebrada', 'Aplazada', 'Cancelada']
ESTADOS_ACUERDO = ['Pendiente', 'En curso', 'Cerrado', 'Descartado']
ESTADOS_PLAN = ['Pendiente', 'En curso', 'Cerrada', 'Descartada']
SITUACIONES = ['Cerrado en plazo', 'Cerrado fuera de plazo', 'Cerrado',
               'VENCIDO', 'Vence esta semana', 'En plazo', 'Descartado']
AREAS = list(DE.AREAS_PLAN_90)
HERRAMIENTAS = list(DE.HERRAMIENTAS_PACK)
assert len(HERRAMIENTAS) == 7


# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto, wrap=True)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, landscape=True):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8


def encabezados(ws, fila, cols, alto=40):
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def bloque(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)


def total(ws, coord, contenido, fmt=None, formula=False):
    cel = (motor.f(ws, coord, contenido, fmt=fmt, bold=True) if formula
           else motor.val(ws, coord, contenido, fmt=fmt, bold=True))
    cel.fill = PatternFill('solid', fgColor=CREMA)
    return cel


def verde_propio(ws, fila, col_et, col_val, etiqueta, valor, fmt, nota,
                 col_nota=None):
    motor.val(ws, col_et + str(fila), etiqueta)
    motor.val(ws, col_val + str(fila), valor, fmt=fmt, verde_=True)
    if col_nota and nota:
        motor.val(ws, col_nota + str(fila), nota, wrap=True)
    return '$' + col_val + '$' + str(fila)


def texto_filas(ws, fila_ini, fila_fin, alto=32):
    for r in range(fila_ini, fila_fin + 1):
        ws.row_dimensions[r].height = alto


def wrap(ws, rango):
    for fila in (ws[rango] if ':' in rango else [[ws[rango]]]):
        for cel in fila:
            cel.alignment = Alignment(vertical='top', wrap_text=True)


def dv_rango(ws, coords, ref, titulo, mensaje):
    """Desplegable cuyo origen es un RANGO de la hoja.

    Necesario para las listas cuyas opciones llevan COMA: `motor.dv_lista`
    las une con comas dentro de una cadena y partiría «Quejas, reclamaciones y
    reseñas» en dos opciones sin dar ningún error.
    """
    dv = DataValidation(type='list', formula1=ref, allow_blank=True,
                        showErrorMessage=True, errorTitle=titulo,
                        error=mensaje)
    ws.add_data_validation(dv)
    for c in coords:
        dv.add(c)
        ws[c].fill = PatternFill('solid', fgColor=motor.VERDE)
        ws[c].protection = openpyxl.styles.Protection(locked=False)
    return dv


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. «Calendario de Reuniones» es tu trimestre: una semanal de equipo cada '
    'dos semanas, una mensual de resultados y un uno-a-uno al mes. Ponlas en '
    'el calendario ANTES de que empiece el trimestre; si no tienen fecha, no '
    'se celebran.',
    '2. «Guion de Reunión Semanal» son siete puntos y treinta minutos. Los '
    'minutos son celdas verdes: si los cambias, el libro te dice si te has '
    'pasado del objetivo. Una reunión de equipo que dura una hora deja de '
    'celebrarse en tres semanas.',
    '3. «Uno-a-uno» trae las seis preguntas y el registro de las doce '
    'personas del equipo. Ninguna pregunta toca la salud, la familia ni la '
    'vida privada: además de ser mala conversación, preguntar por la salud en '
    'una entrevista está prohibido por el art. 9.5 de la Ley 15/2022.',
    '4. «Actas y Acuerdos» es la hoja que hace que las reuniones sirvan de '
    'algo. Cada acuerdo lleva responsable y fecha de seguimiento; la columna '
    '«Situación» te dice sola cuáles están vencidos, cuáles vencen esta semana '
    'y cuáles se cerraron fuera de plazo.',
    '5. Empieza SIEMPRE la reunión semanal por los acuerdos abiertos de la '
    'anterior. Es el punto 4 del guion y es el que convierte la conversación '
    'en trabajo.',
    '6. «Plan 90 Días» recoge las veinte decisiones del trimestre que salen de '
    'las siete herramientas del pack. Escribe el área, de qué libro sale, la '
    'decisión, el responsable y la semana (1-13): la fecha objetivo la calcula '
    'el libro desde la fecha de inicio del plan.',
    '7. Estima el impacto en euros al mes de cada decisión. No hace falta '
    'precisión: hace falta orden de magnitud para saber por dónde empezar y '
    'para poder enseñarle al propietario qué has movido en un trimestre.',
    '8. Si quieres ver los vencimientos «desde» otro día, escribe esa fecha en '
    'la celda verde de simulación de «Actas y Acuerdos». Vacía = hoy.',
]

NOTAS = [
    'ESTA HOJA «PLAN 90 DÍAS» NO ES EL PLAN DE 90 DÍAS DE LA GUÍA FOOD COST. '
    'Aquel plan decide qué pasa con cada PLATO de la carta: escandallar, '
    'resubir, reformular, retirar. Éste ordena lo que sale de las SIETE '
    'herramientas de este manual —personas, servicio, operaciones, '
    'cumplimiento y finanzas— por área, semana y responsable. Se usan EN '
    'PARALELO y no se sustituyen: si tienes los dos, el de la carta te dice '
    'qué platos tocar y éste te dice quién lo hace y cuándo. No los fundas en '
    'uno: acabarías con una lista de treinta cosas sin dueño.',
    'ESTE LIBRO NO TRAE GUION DE BRIEFING DE SERVICIO. El briefing es DIARIO, '
    'dura cinco minutos y ya lo tienes en el Kit de Tareas y en el Kit de '
    'Gestión de Personal. Lo que casi nadie cubre —y es lo que está aquí— son '
    'las otras tres cadencias: la semanal de equipo, la mensual de resultados '
    'y el uno-a-uno.',
    'LOS KPI NO ESTÁN AQUÍ: están en «cuadro-de-mando-semanal-manager.xlsx». '
    'Duplicar el cuadro de mando en la hoja de actas produce dos versiones del '
    'mismo número y, a las tres semanas, discusiones sobre cuál es la buena.',
    'UN ACUERDO SIN RESPONSABLE Y SIN FECHA NO ES UN ACUERDO: ES UNA OPINIÓN. '
    'Por eso las dos columnas están en verde y por eso la «Situación» se queda '
    'vacía si falta la fecha de seguimiento. Si en una reunión no sale ningún '
    'acuerdo con dueño, la reunión no ha terminado.',
    'LA COLUMNA «FECHA DE CIERRE REAL» ESTÁ VACÍA A PROPÓSITO. Anótala cuando '
    'cierres cada acuerdo y el libro te dirá qué porcentaje se cerró DENTRO '
    'del plazo, que es un dato bastante más incómodo (y más útil) que el '
    'porcentaje de cerrados. Mientras no anotes ninguna fecha, ese porcentaje '
    'devuelve vacío: «sin dato» no es «cero».',
    'EL IMPACTO EN EUROS DEL PLAN ES UNA ESTIMACIÓN TUYA, no un cálculo del '
    'libro. La suma sirve para priorizar y para defender el trabajo delante '
    'del propietario, no para presupuestar.',
    'LOS DATOS SEMBRADOS SON UN EJEMPLO MODELADO (el restaurante «La Encina», '
    'el mismo del resto del pack): doce reuniones de un trimestre cerrado, '
    'veinticinco acuerdos —cuatro de ellos vencidos— y veinte decisiones que '
    'salen de las otras seis herramientas. Bórralos y pon los tuyos.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, landscape=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Calendario de Reuniones»
# --------------------------------------------------------------------------
CR0 = 6
CR1 = CR0 + len(DE.REUNIONES) + 11        # 12 reuniones + 12 filas libres


def hoja_reuniones(wb):
    ws = wb.create_sheet('Calendario de Reuniones')
    cabecera(ws, 'Calendario de reuniones — las tres cadencias del trimestre')
    apunte(ws, 'E3',
           'Semanal de equipo, mensual de resultados y uno-a-uno. El briefing '
           'diario no está aquí: es otra cadencia y está en el Kit de Tareas.')

    encabezados(ws, CR0 - 1, [
        ('A', '#', 5), ('B', 'Fecha', 14), ('C', 'Tipo de reunión', 24),
        ('D', 'Cadencia', 12), ('E', 'Asistentes', 26),
        ('F', 'Duración (min)', 12), ('G', 'Responsable', 20),
        ('H', 'Estado', 14), ('I', 'Notas', 52),
    ])
    ws.freeze_panes = 'B' + str(CR0)

    v_tipo, v_cad, v_resp, v_est, v_fecha, v_min = [], [], [], [], [], []
    for i in range(CR1 - CR0 + 1):
        r = CR0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.REUNIONES):
            fecha, tipo, cad, asist, dur, resp, est = DE.REUNIONES[i]
            motor.val(ws, 'B%d' % r, date(*map(int, fecha.split('-'))),
                      fmt=FECHA)
            motor.val(ws, 'C%d' % r, tipo)
            motor.val(ws, 'D%d' % r, cad)
            motor.val(ws, 'E%d' % r, asist)
            motor.val(ws, 'F%d' % r, dur, fmt=ENT)
            motor.val(ws, 'G%d' % r, POR_ID[resp])
            motor.val(ws, 'H%d' % r, est)
        else:
            ws['B%d' % r].number_format = FECHA
            ws['F%d' % r].number_format = ENT
        motor.verde(ws, 'B%d:I%d' % (r, r))
        v_fecha.append('B%d' % r)
        v_tipo.append('C%d' % r)
        v_cad.append('D%d' % r)
        v_min.append('F%d' % r)
        v_resp.append('G%d' % r)
        v_est.append('H%d' % r)
        wrap(ws, 'B%d:I%d' % (r, r))
    texto_filas(ws, CR0, CR1, alto=26)
    motor.dv_fecha(ws, v_fecha)
    motor.dv_lista(ws, v_tipo, TIPOS_REUNION, titulo='Tipo de reunión')
    motor.dv_lista(ws, v_cad, CADENCIAS, titulo='Cadencia')
    motor.dv_lista(ws, v_resp, PERSONAS, titulo='Responsable')
    motor.dv_lista(ws, v_est, ESTADOS_REUNION, titulo='Estado de la reunión')
    motor.dv_numerica(ws, v_min, minimo=5, maximo=240,
                      titulo='Duración (min)',
                      mensaje='Escribe la duración en minutos (5-240).')
    motor.semaforo_texto(ws, 'H%d:H%d' % (CR0, CR1), (
        ('Celebrada', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Prevista', motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('Aplazada', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Cancelada', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))

    B = '$B$%d:$B$%d' % (CR0, CR1)
    C = '$C$%d:$C$%d' % (CR0, CR1)
    F = '$F$%d:$F$%d' % (CR0, CR1)
    H = '$H$%d:$H$%d' % (CR0, CR1)
    fr = CR1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    r = fr + 1
    motor.val(ws, 'B%d' % r, 'Reuniones registradas')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % B, fmt=ENT)
    c_reg = '$D$%d' % r
    r += 1
    for tipo in TIPOS_REUNION:
        motor.val(ws, 'B%d' % r, 'De tipo «%s»' % tipo)
        motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"%s")' % (C, tipo), fmt=ENT)
        r += 1
    motor.val(ws, 'B%d' % r, 'Celebradas')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"Celebrada")' % H, fmt=ENT)
    c_celeb = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Celebradas (%)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_reg, c_celeb, c_reg),
            fmt=PCT, bold=True)
    c_pct = '$D$%d' % r
    r += 1
    total(ws, 'B%d' % r, 'MINUTOS DE REUNIÓN DEL TRIMESTRE')
    total(ws, 'D%d' % r, '=IFERROR(SUM(%s),"")' % F, fmt=ENT, formula=True)
    c_min = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Duración media (min)')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_reg, c_min, c_reg),
            fmt='#,##0.0')
    r += 1
    motor.val(ws, 'B%d' % r, 'Horas de reunión del trimestre')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s="","",%s/60),"")' % (c_min, c_min), fmt='#,##0.0')
    r += 2
    p_obj = verde_propio(
        ws, r, 'B', 'D', 'Objetivo de reuniones celebradas (%)', 0.90, PCT,
        'Por debajo de este porcentaje, la celda «Celebradas (%)» se pone en '
        'rojo. Una reunión que se cancela dos veces seguidas deja de existir.',
        col_nota='F')
    motor.dv_porcentaje(ws, ['D%d' % r], titulo='Objetivo de celebradas',
                        prompt='Se escribe en tanto por uno: 0,90 = 90 %.')
    motor.regla_expresion(ws, c_pct.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<%s)' % (c_pct, c_pct, p_obj))
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Guion de Reunión Semanal»
# --------------------------------------------------------------------------
GR0 = 6
GR1 = GR0 + len(DE.GUION_REUNION_SEMANAL) + 2   # 7 puntos + 3 filas libres

# El responsable de cada punto no está en `datos_ejemplo.py` (que sólo fija
# orden, punto, minutos y herramienta): se asigna por el ROL de la plantilla.
# El punto 6 es del equipo por turno rotatorio a propósito.
RESP_GUION = {1: 'P01', 2: 'P01', 3: 'P03', 4: 'P01', 5: 'P01',
              6: None, 7: 'P01'}
POR_QUE = {
    1: 'Ventas, prime cost y cubiertos de la semana que cierra. Cinco minutos '
       'de números, no veinte: el detalle está en el cuadro de mando.',
    2: 'Un semáforo en rojo sin explicación se repite. Aquí se nombra la causa '
       'y se decide si hace falta un acuerdo.',
    3: 'Lo que se repite en las quejas y en las reseñas es un problema de '
       'PROCESO, no de la persona que estaba ese día.',
    4: 'El punto que hace que la reunión sirva de algo: se leen los acuerdos '
       'abiertos de la semana pasada, uno por uno, con su responsable delante.',
    # M8 (auditoría 2026-09-04): «cuadrante» es vocabulario de España; en
    # buena parte de LATAM se dice «rol» u «horario». Primera aparición del
    # término en este libro.
    5: 'Reservas grandes, eventos, cambios de cuadrante (rol u horario, en '
       'el uso de otros países) y quién falta. Es la única parte de la '
       'reunión que mira hacia delante.',
    6: 'Un punto que trae el equipo, por turno rotatorio. Se anuncia en la '
       'reunión anterior para que venga preparado.',
    7: 'Qué, quién y para cuándo. Se escriben en la hoja «Actas y Acuerdos» '
       'antes de levantarse de la mesa, no después.',
}


def hoja_guion(wb):
    ws = wb.create_sheet('Guion de Reunión Semanal')
    cabecera(ws, 'Guion de la reunión semanal — siete puntos, treinta minutos')
    apunte(ws, 'D3',
           'El orden importa: los números primero y los acuerdos al final, '
           'para que nadie se marche antes de que se reparta el trabajo.')

    encabezados(ws, GR0 - 1, [
        ('A', 'Orden', 7), ('B', 'Punto', 50), ('C', 'Minutos', 10),
        ('D', 'Herramienta de la que salen los datos', 34),
        ('E', 'Responsable', 20), ('F', 'Por qué está en el guion', 62),
        ('G', 'Notas', 40),
    ])
    ws.freeze_panes = 'B' + str(GR0)

    v_min, v_resp = [], []
    for i in range(GR1 - GR0 + 1):
        r = GR0 + i
        if i < len(DE.GUION_REUNION_SEMANAL):
            orden, punto, minutos, herramienta = DE.GUION_REUNION_SEMANAL[i]
            motor.val(ws, 'A%d' % r, orden, fmt=ENT)
            motor.val(ws, 'B%d' % r, punto, bold=True)
            motor.val(ws, 'C%d' % r, minutos, fmt=ENT)
            motor.val(ws, 'D%d' % r, herramienta)
            rid = RESP_GUION[orden]
            motor.val(ws, 'E%d' % r,
                      POR_ID[rid] if rid else 'Turno rotatorio del equipo')
            motor.val(ws, 'F%d' % r, POR_QUE[orden])
        else:
            motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
            ws['C%d' % r].number_format = ENT
        motor.verde(ws, 'B%d:E%d' % (r, r))
        motor.verde(ws, 'G%d' % r)
        v_min.append('C%d' % r)
        v_resp.append('E%d' % r)
        wrap(ws, 'B%d:G%d' % (r, r))
    texto_filas(ws, GR0, GR1, alto=46)
    motor.dv_numerica(ws, v_min, minimo=0, maximo=120, titulo='Minutos',
                      mensaje='Escribe los minutos del punto (0-120).')
    motor.dv_lista(ws, v_resp, PERSONAS + ['Turno rotatorio del equipo'],
                   titulo='Responsable del punto')

    C = '$C$%d:$C$%d' % (GR0, GR1)
    B = '$B$%d:$B$%d' % (GR0, GR1)
    fr = GR1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    r = fr + 1
    motor.val(ws, 'B%d' % r, 'Puntos del guion')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % B, fmt=ENT)
    r += 1
    total(ws, 'B%d' % r, 'DURACIÓN TOTAL DEL GUION (min)')
    total(ws, 'D%d' % r, '=IFERROR(SUM(%s),"")' % C, fmt=ENT, formula=True)
    c_total = '$D$%d' % r
    r += 1
    p_obj = verde_propio(
        ws, r, 'B', 'D', 'Duración objetivo de la reunión (min)', 30, ENT,
        'Si la suma de los minutos se pasa de aquí, la celda de abajo se pone '
        'en rojo. Treinta minutos de pie funcionan mejor que una hora '
        'sentados.', col_nota='F')
    motor.dv_numerica(ws, ['D%d' % r], minimo=5, maximo=240,
                      titulo='Duración objetivo',
                      mensaje='Escribe la duración objetivo en minutos '
                              '(5-240).')
    r += 1
    motor.val(ws, 'B%d' % r, 'Diferencia con el objetivo (min)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(OR(%s="",%s=""),"",%s-%s),"")'
            % (c_total, p_obj, c_total, p_obj), fmt=ENT, bold=True)
    c_dif = '$D$%d' % r
    motor.regla_expresion(ws, c_dif.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s>0)' % (c_dif, c_dif))
    r += 2
    apunte(ws, 'B%d' % r,
           'El punto 6 lo trae el equipo por turno rotatorio y se anuncia en '
           'la reunión anterior. Es lo único de esta lista que no controla el '
           'manager, y es lo primero que se cae cuando hay prisa: no lo dejes '
           'caer.')
    ws.row_dimensions[r].height = 30
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Uno-a-uno»
# --------------------------------------------------------------------------
U_P0 = 6                                  # preguntas
U_P1 = U_P0 + len(DE.UNO_A_UNO) - 1
U_S0 = U_P1 + 4                           # cabecera del registro
U_S1 = U_S0 + len(DE.PLANTILLA)           # 12 sesiones

# Las tres sesiones ya celebradas del trimestre salen del cruce entre
# REUNIONES (los tres uno-a-uno y con quién) y ACUERDOS (lo que se acordó ese
# mismo día con esa persona). No se inventa ninguna.
SESIONES = {
    'P08': ('2026-06-15', 'Reparto de responsabilidad en el cierre de caja',
            'A07', '2026-07-01', 'Cerrado'),
    'P05': ('2026-07-20', 'Partida fría: es el único punto único de fallo del '
                          'equipo', 'A15', '2026-10-15', 'En curso'),
    'P09': ('2026-08-17', 'Ruta profesional hacia jefa de rango',
            'A24', '2026-08-31', 'Pendiente'),
}
ACUERDO_TXT = {a[0]: a[2] for a in DE.ACUERDOS}
ESTADOS_SESION = ['Pendiente', 'En curso', 'Cerrado']


def hoja_uno_a_uno(wb):
    ws = wb.create_sheet('Uno-a-uno')
    cabecera(ws, 'Uno-a-uno — seis preguntas y el registro del equipo')
    apunte(ws, 'D3',
           'Ninguna de las seis preguntas toca la salud, la familia ni la vida '
           'privada. No es sólo buena educación: preguntarlo está prohibido '
           '(art. 9.5 de la Ley 15/2022).')

    encabezados(ws, U_P0 - 1, [
        ('A', '#', 5), ('B', 'Pregunta', 62),
        ('C', 'Para qué sirve', 62), ('D', 'Notas', 44),
    ])
    PARA_QUE = [
        'Abre en positivo y saca a la luz lo que está funcionando y no se '
        'cuenta nunca.',
        'La pregunta más útil de las seis: convierte una queja en una petición '
        'concreta que tú puedes resolver.',
        'El que está en la partida ve cosas que tú no ves desde la oficina. '
        'Aquí es donde salen.',
        'Es la que alimenta la matriz de formación y polivalencia: quién '
        'quiere aprender qué.',
        'La pregunta abierta obligatoria. Deja el silencio: la respuesta suele '
        'llegar después de unos segundos.',
        'Cierra con un compromiso de cada lado. Lo que salga de aquí va a la '
        'hoja «Actas y Acuerdos», no a tu memoria.',
    ]
    for i, pregunta in enumerate(DE.UNO_A_UNO):
        r = U_P0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, pregunta, bold=True)
        motor.val(ws, 'C%d' % r, PARA_QUE[i])
        motor.verde(ws, 'D%d' % r)
        wrap(ws, 'B%d:D%d' % (r, r))
    texto_filas(ws, U_P0, U_P1, alto=34)

    bloque(ws, 'A%d' % (U_S0 - 2),
           'REGISTRO DE SESIONES — una fila por persona del equipo')
    encabezados(ws, U_S0 - 1, [
        ('A', '#', 5), ('B', 'Empleado', 20), ('C', 'Fecha', 14),
        ('D', 'Tema principal', 44), ('E', 'Acuerdo', 52),
        ('F', 'Fecha de seguimiento', 14), ('G', 'Estado', 14),
        ('H', 'Notas', 40),
    ])
    v_emp, v_f1, v_f2, v_est = [], [], [], []
    for i, p in enumerate(DE.PLANTILLA):
        r = U_S0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        motor.val(ws, 'B%d' % r, POR_ID[p[0]])
        if p[0] in SESIONES:
            fecha, tema, aid, seg, est = SESIONES[p[0]]
            motor.val(ws, 'C%d' % r, date(*map(int, fecha.split('-'))),
                      fmt=FECHA)
            motor.val(ws, 'D%d' % r, tema)
            motor.val(ws, 'E%d' % r, ACUERDO_TXT[aid] + ' (acuerdo ' + aid + ')')
            motor.val(ws, 'F%d' % r, date(*map(int, seg.split('-'))), fmt=FECHA)
            motor.val(ws, 'G%d' % r, est)
        else:
            ws['C%d' % r].number_format = FECHA
            ws['F%d' % r].number_format = FECHA
        motor.verde(ws, 'B%d:H%d' % (r, r))
        v_emp.append('B%d' % r)
        v_f1.append('C%d' % r)
        v_f2.append('F%d' % r)
        v_est.append('G%d' % r)
        wrap(ws, 'B%d:H%d' % (r, r))
    texto_filas(ws, U_S0, U_S1, alto=32)
    motor.dv_lista(ws, v_emp, PERSONAS, titulo='Empleado')
    motor.dv_fecha(ws, v_f1 + v_f2)
    motor.dv_lista(ws, v_est, ESTADOS_SESION, titulo='Estado del acuerdo')
    motor.semaforo_texto(ws, 'G%d:G%d' % (U_S0, U_S1), (
        ('Cerrado', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('En curso', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Pendiente', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    B = '$B$%d:$B$%d' % (U_S0, U_S1)
    C = '$C$%d:$C$%d' % (U_S0, U_S1)
    fr = U_S1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    r = fr + 1
    motor.val(ws, 'B%d' % r, 'Personas del equipo en el registro')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % B, fmt=ENT)
    c_pers = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Sesiones con fecha anotada')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % C, fmt=ENT)
    c_ses = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Personas sin uno-a-uno todavía')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s="","",%s-%s),"")' % (c_pers, c_pers, c_ses),
            fmt=ENT)
    r += 1
    total(ws, 'B%d' % r, 'EQUIPO CON UNO-A-UNO HECHO (%)')
    total(ws, 'D%d' % r,
          '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_pers, c_ses, c_pers),
          fmt=PCT, formula=True)
    c_pct = '$D$%d' % r
    r += 1
    p_obj = verde_propio(
        ws, r, 'B', 'D', 'Objetivo de cobertura del equipo (%)', 1.00, PCT,
        'La semana 8 del plan de 90 días dice «uno-a-uno con las doce personas '
        'del equipo, uno por semana». Ése es el objetivo.', col_nota='F')
    motor.dv_porcentaje(ws, ['D%d' % r], titulo='Objetivo de cobertura',
                        prompt='Se escribe en tanto por uno: 1,00 = 100 %.')
    motor.regla_expresion(ws, c_pct.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<%s)' % (c_pct, c_pct, p_obj))
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Actas y Acuerdos»
# --------------------------------------------------------------------------
AA0 = 9
AA1 = AA0 + len(DE.ACUERDOS) + 9          # 25 acuerdos + 10 filas libres
SIM_A, HOY_A, VENTANA = '$C$4', '$C$5', '$C$6'
TIPO_POR_FECHA = {r[0]: r[1] for r in DE.REUNIONES}


def hoja_acuerdos(wb):
    ws = wb.create_sheet('Actas y Acuerdos')
    cabecera(ws, 'Actas y acuerdos — lo que la reunión deja escrito')
    apunte(ws, 'F3',
           'Un acuerdo sin responsable y sin fecha no es un acuerdo: es una '
           'opinión. La columna «Situación» la calcula el libro.')

    motor.val(ws, 'B4', 'Fecha para simular (déjala vacía y se usa la de hoy)')
    motor.val(ws, 'C4', None, fmt=FECHA, verde_=True)
    motor.dv_fecha(ws, ['C4'])
    motor.val(ws, 'D4',
              'Escribe aquí una fecha para ver los vencimientos «desde» ese '
              'día. Vacía = hoy.', wrap=True)
    motor.val(ws, 'B5', 'Hoy (fecha de referencia del libro)', bold=True)
    motor.f(ws, 'C5', '=IFERROR(IF(%s="",TODAY(),%s),"")' % (SIM_A, SIM_A),
            fmt=FECHA, bold=True)
    motor.val(ws, 'D5',
              'Es la única celda de todo el libro que usa la fecha del '
              'sistema. La hoja «Plan 90 Días» la lee de aquí.', wrap=True)
    motor.val(ws, 'B6', 'Días que cuentan como «vence esta semana»')
    motor.val(ws, 'C6', 7, fmt=ENT, verde_=True)
    motor.dv_numerica(ws, ['C6'], minimo=1, maximo=60,
                      titulo='Ventana de aviso',
                      mensaje='Escribe cuántos días por delante quieres que '
                              'el libro avise (1-60).')

    encabezados(ws, AA0 - 1, [
        ('A', 'Id', 6), ('B', 'Fecha de la reunión', 14),
        ('C', 'Tipo de reunión', 22), ('D', 'Acuerdo o decisión', 54),
        ('E', 'Responsable', 20), ('F', 'Fecha de seguimiento', 14),
        ('G', 'Estado', 14), ('H', 'Fecha de cierre real', 14),
        ('I', 'Situación', 20), ('J', 'Días hasta el seguimiento', 12),
        ('K', 'Notas', 44),
    ])
    ws.freeze_panes = 'B' + str(AA0)

    v_fecha, v_tipo, v_resp, v_est = [], [], [], []
    for i in range(AA1 - AA0 + 1):
        r = AA0 + i
        if i < len(DE.ACUERDOS):
            aid, freu, acuerdo, resp, seg, est, cierre = DE.ACUERDOS[i]
            motor.val(ws, 'A%d' % r, aid)
            motor.val(ws, 'B%d' % r, date(*map(int, freu.split('-'))),
                      fmt=FECHA)
            motor.val(ws, 'C%d' % r, TIPO_POR_FECHA[freu])
            motor.val(ws, 'D%d' % r, acuerdo)
            motor.val(ws, 'E%d' % r, POR_ID[resp])
            motor.val(ws, 'F%d' % r, date(*map(int, seg.split('-'))),
                      fmt=FECHA)
            motor.val(ws, 'G%d' % r, est)
            if cierre:
                motor.val(ws, 'H%d' % r, date(*map(int, cierre.split('-'))),
                          fmt=FECHA)
            else:
                ws['H%d' % r].number_format = FECHA
        else:
            motor.val(ws, 'A%d' % r, 'A%02d' % (i + 1))
            ws['B%d' % r].number_format = FECHA
            ws['F%d' % r].number_format = FECHA
        ws['H%d' % r].number_format = FECHA
        motor.verde(ws, 'B%d:H%d' % (r, r))
        motor.verde(ws, 'K%d' % r)
        v_fecha += ['B%d' % r, 'F%d' % r, 'H%d' % r]
        v_tipo.append('C%d' % r)
        v_resp.append('E%d' % r)
        v_est.append('G%d' % r)
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF(OR($F{r}="",{hoy}=""),"",$F{r}-{hoy}),"")'
                .format(r=r, hoy=HOY_A), fmt=ENT)
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF($D{r}="","",'
                'IF($G{r}="Descartado","Descartado",'
                'IF($G{r}="Cerrado",IF($H{r}="","Cerrado",'
                'IF($H{r}<=$F{r},"Cerrado en plazo","Cerrado fuera de plazo")),'
                'IF($F{r}="","",IF($J{r}<0,"VENCIDO",'
                'IF($J{r}<={v},"Vence esta semana","En plazo")))))),"")'
                .format(r=r, v=VENTANA))
        wrap(ws, 'B%d:K%d' % (r, r))
    texto_filas(ws, AA0, AA1, alto=30)
    motor.dv_fecha(ws, v_fecha)
    motor.dv_lista(ws, v_tipo, TIPOS_REUNION, titulo='Tipo de reunión')
    motor.dv_lista(ws, v_resp, PERSONAS, titulo='Responsable')
    motor.dv_lista(ws, v_est, ESTADOS_ACUERDO, titulo='Estado del acuerdo')
    motor.semaforo_texto(ws, 'G%d:G%d' % (AA0, AA1), (
        ('Cerrado', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('En curso', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Pendiente', motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('Descartado', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))
    motor.semaforo_texto(ws, 'I%d:I%d' % (AA0, AA1), (
        ('VENCIDO', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('Cerrado fuera de plazo', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('Vence esta semana', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Cerrado en plazo', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Cerrado', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('En plazo', motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('Descartado', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))
    motor.semaforo_isnumber(ws, 'J%d:J%d' % (AA0, AA1), '$J%d' % AA0,
                            operador='<', umbral='0')

    D = '$D$%d:$D$%d' % (AA0, AA1)
    E = '$E$%d:$E$%d' % (AA0, AA1)
    G = '$G$%d:$G$%d' % (AA0, AA1)
    H = '$H$%d:$H$%d' % (AA0, AA1)
    I = '$I$%d:$I$%d' % (AA0, AA1)
    fr = AA1 + 2
    bloque(ws, 'A%d' % fr, 'RESUMEN — lo calcula el libro')
    r = fr + 1
    motor.val(ws, 'B%d' % r, 'Acuerdos registrados')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % D, fmt=ENT)
    c_reg = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Cerrados')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"Cerrado")' % G, fmt=ENT)
    c_cerr = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Descartados')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"Descartado")' % G, fmt=ENT)
    c_desc = '$D$%d' % r
    r += 1
    total(ws, 'B%d' % r, 'ACUERDOS ABIERTOS')
    total(ws, 'D%d' % r,
          '=IFERROR(IF(%s="","",%s-%s-%s),"")' % (c_reg, c_reg, c_cerr,
                                                  c_desc),
          fmt=ENT, formula=True)
    c_abie = '$D$%d' % r
    r += 1
    total(ws, 'B%d' % r, 'ACUERDOS VENCIDOS')
    total(ws, 'D%d' % r, '=COUNTIF(%s,"VENCIDO")' % I, fmt=ENT, formula=True)
    c_venc = '$D$%d' % r
    motor.regla_expresion(ws, c_venc.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s>0)' % (c_venc, c_venc))
    r += 1
    motor.val(ws, 'B%d' % r, 'Vencen esta semana')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"Vence esta semana")' % I, fmt=ENT)
    r += 1
    motor.val(ws, 'B%d' % r, 'Acuerdos cerrados (%)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_reg, c_cerr, c_reg),
            fmt=PCT, bold=True)
    c_pctc = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Acuerdos en plazo (%)')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",(%s-%s)/%s),"")'
            % (c_reg, c_reg, c_venc, c_reg), fmt=PCT)
    r += 1
    motor.val(ws, 'B%d' % r, 'Cerrados con fecha de cierre anotada')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % H, fmt=ENT)
    c_con_fecha = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Cerrados EN PLAZO (%)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",COUNTIF(%s,"Cerrado en plazo")/%s),"")'
            % (c_con_fecha, I, c_con_fecha), fmt=PCT, bold=True)
    r += 1
    motor.val(ws, 'D%d' % r,
              'Vacío mientras no anotes ninguna fecha de cierre real: «sin '
              'dato» no es «cero».')
    ws['D%d' % r].font = Font(size=9, color=GRIS)
    r += 2
    p_obj = verde_propio(
        ws, r, 'B', 'D', 'Objetivo de acuerdos en plazo (%)', 0.90, PCT,
        'Por debajo de este porcentaje la celda «Acuerdos en plazo (%)» se '
        'pone en rojo.', col_nota='F')
    motor.dv_porcentaje(ws, ['D%d' % r], titulo='Objetivo en plazo',
                        prompt='Se escribe en tanto por uno: 0,90 = 90 %.')
    c_pcte = '$D$%d' % (r - 5)
    motor.regla_expresion(ws, c_pcte.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<%s)' % (c_pcte, c_pcte, p_obj))

    # --- por responsable ---------------------------------------------------
    rr = r + 2
    bloque(ws, 'A%d' % rr, 'ACUERDOS POR RESPONSABLE — lo calcula el libro')
    encabezados(ws, rr + 1, [
        ('B', 'Responsable', None), ('C', 'Acuerdos', None),
        ('D', 'Abiertos', None), ('E', 'Vencidos', None),
    ], alto=32)
    for k, persona in enumerate(PERSONAS):
        f = rr + 2 + k
        motor.val(ws, 'B%d' % f, persona, bold=True)
        motor.f(ws, 'C%d' % f, '=COUNTIF(%s,$B%d)' % (E, f), fmt=ENT)
        motor.f(ws, 'D%d' % f,
                '=IFERROR(SUMPRODUCT(--(%s=$B%d),--(%s<>"Cerrado"),'
                '--(%s<>"Descartado"),--(%s<>"")),"")' % (E, f, G, G, G),
                fmt=ENT)
        motor.f(ws, 'E%d' % f,
                '=IFERROR(SUMPRODUCT(--(%s=$B%d),--(%s="VENCIDO")),"")'
                % (E, f, I), fmt=ENT)
    setup(ws)
    return ws, rr + 2


# --------------------------------------------------------------------------
# Hoja «Plan 90 Días»
# --------------------------------------------------------------------------
PL0 = 6
PL1 = PL0 + len(DE.PLAN_90) + 7           # 20 decisiones + 8 filas libres


def hoja_plan(wb):
    ws = wb.create_sheet('Plan 90 Días')
    cabecera(ws, 'Plan de 90 días — lo que sale de las siete herramientas')
    apunte(ws, 'F3',
           'NO es el plan de 90 días de la Guía Food Cost (aquel decide qué '
           'pasa con cada plato). Éste ordena lo que sale de este manual. Se '
           'usan en paralelo.')

    encabezados(ws, PL0 - 1, [
        ('A', '#', 5), ('B', 'Área', 16),
        ('C', 'Herramienta de origen', 30), ('D', 'Decisión', 56),
        ('E', 'Responsable', 20), ('F', 'Semana (1-13)', 10),
        ('G', 'Fecha objetivo', 14), ('H', 'Impacto estimado (€/mes)', 14),
        ('I', 'Estado', 14), ('J', 'Impacto ya conseguido (€/mes)', 14),
        ('K', 'Notas', 40),
    ])
    ws.freeze_panes = 'B' + str(PL0)

    # Parámetros del plan (bloque de resumen), referenciados desde las filas.
    fr = PL1 + 2
    f_inicio = '$D$%d' % (fr + 12)
    f_semanas = '$D$%d' % (fr + 13)

    v_area, v_her, v_resp, v_est, v_sem, v_eur = [], [], [], [], [], []
    for i in range(PL1 - PL0 + 1):
        r = PL0 + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.PLAN_90):
            area, herram, decision, resp, semana, impacto, estado = \
                DE.PLAN_90[i]
            motor.val(ws, 'B%d' % r, area)
            motor.val(ws, 'C%d' % r, herram)
            motor.val(ws, 'D%d' % r, decision)
            motor.val(ws, 'E%d' % r, POR_ID[resp])
            motor.val(ws, 'F%d' % r, semana, fmt=ENT)
            motor.val(ws, 'H%d' % r, impacto, fmt=EUR)
            motor.val(ws, 'I%d' % r, estado)
        else:
            ws['F%d' % r].number_format = ENT
            ws['H%d' % r].number_format = EUR
        motor.verde(ws, 'B%d:F%d' % (r, r))
        motor.verde(ws, 'H%d:I%d' % (r, r))
        motor.verde(ws, 'K%d' % r)
        v_area.append('B%d' % r)
        v_her.append('C%d' % r)
        v_resp.append('E%d' % r)
        v_est.append('I%d' % r)
        v_sem.append('F%d' % r)
        v_eur.append('H%d' % r)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF(OR($F{r}="",{ini}=""),"",{ini}+7*($F{r}-1)),"")'
                .format(r=r, ini=f_inicio), fmt=FECHA)
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF(OR($D{r}="",$H{r}=""),"",'
                'IF($I{r}="Cerrada",$H{r},0)),"")'.format(r=r), fmt=EUR)
        ws['J%d' % r].font = Font(bold=True, color=AZUL)
        wrap(ws, 'B%d:K%d' % (r, r))
    texto_filas(ws, PL0, PL1, alto=34)

    motor.dv_lista(ws, v_area, AREAS, titulo='Área del plan')
    motor.dv_lista(ws, v_resp, PERSONAS, titulo='Responsable')
    motor.dv_lista(ws, v_est, ESTADOS_PLAN, titulo='Estado de la decisión')
    motor.dv_numerica(ws, v_sem, minimo=1, maximo=13,
                      titulo='Semana del plan',
                      mensaje='El plan tiene 13 semanas: escribe un número '
                              'entre 1 y 13.')
    motor.dv_numerica(ws, v_eur, minimo=0, titulo='Impacto estimado',
                      mensaje='Escribe el impacto estimado en euros al mes '
                              '(0 o más).')
    motor.semaforo_texto(ws, 'I%d:I%d' % (PL0, PL1), (
        ('Cerrada', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('En curso', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
        ('Pendiente', motor.CF_GRIS_BG, motor.CF_GRIS_FG),
        ('Descartada', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))

    B = '$B$%d:$B$%d' % (PL0, PL1)
    C = '$C$%d:$C$%d' % (PL0, PL1)
    D = '$D$%d:$D$%d' % (PL0, PL1)
    H = '$H$%d:$H$%d' % (PL0, PL1)
    I = '$I$%d:$I$%d' % (PL0, PL1)
    J = '$J$%d:$J$%d' % (PL0, PL1)

    bloque(ws, 'A%d' % fr, 'RESUMEN Y PARÁMETROS — lo calcula el libro')
    r = fr + 1
    motor.val(ws, 'B%d' % r, 'Decisiones registradas')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"<>")' % D, fmt=ENT)
    c_reg = '$D$%d' % r
    r += 1
    for etiqueta, criterio in [('En curso', 'En curso'),
                               ('Pendientes', 'Pendiente'),
                               ('Descartadas', 'Descartada')]:
        motor.val(ws, 'B%d' % r, etiqueta)
        motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"%s")' % (I, criterio), fmt=ENT)
        r += 1
    motor.val(ws, 'B%d' % r, 'Cerradas')
    motor.f(ws, 'D%d' % r, '=COUNTIF(%s,"Cerrada")' % I, fmt=ENT)
    c_cerr = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Decisiones cerradas (%)', bold=True)
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(%s=0,"",%s/%s),"")' % (c_reg, c_cerr, c_reg),
            fmt=PCT, bold=True)
    c_pct = '$D$%d' % r
    r += 1
    total(ws, 'B%d' % r, 'IMPACTO TOTAL ESTIMADO (€/mes)')
    total(ws, 'D%d' % r, '=IFERROR(SUM(%s),"")' % H, fmt=EUR, formula=True)
    c_imp = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Impacto ya conseguido (€/mes)')
    motor.f(ws, 'D%d' % r, '=IFERROR(SUM(%s),"")' % J, fmt=EUR)
    c_con = '$D$%d' % r
    r += 1
    motor.val(ws, 'B%d' % r, 'Impacto pendiente (€/mes)')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(OR(%s="",%s=""),"",%s-%s),"")'
            % (c_imp, c_con, c_imp, c_con), fmt=EUR)
    r += 1
    motor.val(ws, 'B%d' % r, 'Impacto total estimado a 12 meses (€)')
    motor.f(ws, 'D%d' % r, '=IFERROR(IF(%s="","",%s*12),"")' % (c_imp, c_imp),
            fmt=EUR)
    r += 1
    # --- parámetros (verdes) ---------------------------------------------
    assert r + 1 == fr + 12, 'La fila de la fecha de inicio se ha movido'
    r += 1
    p_ini = verde_propio(
        ws, r, 'B', 'D', 'Fecha de inicio del plan (lunes de la semana 1)',
        date(*map(int, DE.RESTAURANTE['fecha_inicio_plan_90'].split('-'))),
        FECHA,
        'La semana 1 empieza este día. De aquí salen todas las fechas '
        'objetivo. Cámbiala por la tuya.', col_nota='F')
    motor.dv_fecha(ws, ['D%d' % r])
    assert p_ini == f_inicio, 'p_ini %s != %s' % (p_ini, f_inicio)
    r += 1
    p_sem = verde_propio(
        ws, r, 'B', 'D', 'Semanas del plan', 13, ENT,
        'Trece semanas son noventa días. Si tu trimestre es otro, cámbialo.',
        col_nota='F')
    motor.dv_numerica(ws, ['D%d' % r], minimo=1, maximo=52,
                      titulo='Semanas del plan',
                      mensaje='Escribe cuántas semanas dura el plan (1-52).')
    assert p_sem == f_semanas, 'p_sem %s != %s' % (p_sem, f_semanas)
    r += 1
    motor.val(ws, 'B%d' % r, 'Fecha de cierre del plan')
    motor.f(ws, 'D%d' % r,
            '=IFERROR(IF(OR(%s="",%s=""),"",%s+7*%s-1),"")'
            % (p_ini, p_sem, p_ini, p_sem), fmt=FECHA)
    r += 1
    motor.val(ws, 'B%d' % r, 'Semana en curso del plan')
    motor.f(ws, 'D%d' % r,
            "=IFERROR(IF(OR(%s=\"\",%s=\"\",'Actas y Acuerdos'!%s=\"\"),\"\","
            "IF('Actas y Acuerdos'!%s<%s,\"\","
            "MIN(%s,INT(('Actas y Acuerdos'!%s-%s)/7)+1))),\"\")"
            % (p_ini, p_sem, HOY_A, HOY_A, p_ini, p_sem, HOY_A, p_ini),
            fmt=ENT)
    r += 1
    motor.val(ws, 'B%d' % r, 'Avance del calendario (%)', bold=True)
    motor.f(ws, 'D%d' % r,
            "=IFERROR(IF(OR(%s=\"\",%s=\"\",'Actas y Acuerdos'!%s=\"\"),\"\","
            "MAX(0,MIN(1,('Actas y Acuerdos'!%s-%s)/(%s*7)))),\"\")"
            % (p_ini, p_sem, HOY_A, HOY_A, p_ini, p_sem), fmt=PCT, bold=True)
    r += 1
    p_obj = verde_propio(
        ws, r, 'B', 'D', 'Objetivo de decisiones cerradas a 90 días (%)', 0.80,
        PCT,
        'Qué parte de las decisiones te comprometes a cerrar en el trimestre. '
        'Si el porcentaje cerrado se queda por debajo, la celda se pone en '
        'rojo.', col_nota='F')
    motor.dv_porcentaje(ws, ['D%d' % r], titulo='Objetivo de cierre',
                        prompt='Se escribe en tanto por uno: 0,80 = 80 %.')
    motor.regla_expresion(ws, c_pct.replace('$', ''),
                          '=AND(ISNUMBER(%s),%s<%s)' % (c_pct, c_pct, p_obj))

    # --- resúmenes por área y por herramienta -----------------------------
    ra = r + 2
    bloque(ws, 'A%d' % ra, 'RESUMEN POR ÁREA — lo calcula el libro')
    encabezados(ws, ra + 1, [
        ('B', 'Área', None), ('C', 'Decisiones', None),
        ('D', 'Impacto estimado (€/mes)', None),
        ('E', 'Impacto conseguido (€/mes)', None),
    ], alto=32)
    for k, area in enumerate(AREAS):
        f = ra + 2 + k
        motor.val(ws, 'B%d' % f, area, bold=True)
        motor.f(ws, 'C%d' % f, '=COUNTIF(%s,$B%d)' % (B, f), fmt=ENT)
        motor.f(ws, 'D%d' % f,
                '=IFERROR(SUMIF(%s,$B%d,%s),"")' % (B, f, H), fmt=EUR)
        motor.f(ws, 'E%d' % f,
                '=IFERROR(SUMPRODUCT(--(%s=$B%d),--(%s="Cerrada"),%s),"")'
                % (B, f, I, H), fmt=EUR)
    fila_area = ra + 2

    rh = ra + 2 + len(AREAS) + 1
    bloque(ws, 'A%d' % rh,
           'RESUMEN POR HERRAMIENTA DE ORIGEN — lo calcula el libro')
    encabezados(ws, rh + 1, [
        ('B', 'Herramienta de origen', None), ('C', 'Decisiones', None),
        ('D', 'Impacto estimado (€/mes)', None),
        ('E', 'Impacto conseguido (€/mes)', None),
    ], alto=32)
    for k, herram in enumerate(HERRAMIENTAS):
        f = rh + 2 + k
        motor.val(ws, 'B%d' % f, herram, bold=True)
        motor.f(ws, 'C%d' % f, '=COUNTIF(%s,$B%d)' % (C, f), fmt=ENT)
        motor.f(ws, 'D%d' % f,
                '=IFERROR(SUMIF(%s,$B%d,%s),"")' % (C, f, H), fmt=EUR)
        motor.f(ws, 'E%d' % f,
                '=IFERROR(SUMPRODUCT(--(%s=$B%d),--(%s="Cerrada"),%s),"")'
                % (C, f, I, H), fmt=EUR)
    fila_herr = rh + 2

    # Los siete nombres de herramienta llevan COMA: la DV va contra el rango
    # del propio resumen, no contra una cadena separada por comas.
    dv_rango(ws, v_her,
             "'Plan 90 Días'!$B$%d:$B$%d" % (fila_herr,
                                             fila_herr + len(HERRAMIENTAS) - 1),
             'Herramienta no válida',
             'Elige una de las siete herramientas del pack (la lista está en '
             'el bloque «Resumen por herramienta de origen» de esta misma '
             'hoja).')

    fin = fila_herr + len(HERRAMIENTAS) + 1
    apunte(ws, 'B%d' % fin,
           'A los 90 días, vuelve a las siete herramientas con los datos '
           'nuevos: el cuadro de mando semanal te dirá si el prime cost se '
           'movió, la matriz si el punto único de fallo sigue ahí y el '
           'calendario legal si queda algo vencido. Y se abre el trimestre '
           'siguiente.')
    ws.row_dimensions[fin].height = 44
    setup(ws)
    return ws, fila_area, fila_herr


# --------------------------------------------------------------------------
def mapa(fila_resp, fila_area, fila_herr):
    crr = CR1 + 3
    grr = GR1 + 3
    urr = U_S1 + 3
    arr = AA1 + 3
    prr = PL1 + 3
    cel_reu = {'Reuniones registradas': 'D%d' % crr}
    for k, tipo in enumerate(TIPOS_REUNION):
        cel_reu['Reuniones de tipo «%s»' % tipo] = 'D%d' % (crr + 1 + k)
    cel_reu.update({
        'Celebradas': 'D%d' % (crr + 4),
        'Celebradas (%)': 'D%d' % (crr + 5),
        'MINUTOS DE REUNIÓN DEL TRIMESTRE': 'D%d' % (crr + 6),
        'Duración media (min)': 'D%d' % (crr + 7),
        'Horas de reunión del trimestre': 'D%d' % (crr + 8),
        'Objetivo de reuniones celebradas (%)': 'D%d' % (crr + 10),
    })
    cel_resp = {}
    for k, persona in enumerate(PERSONAS):
        cel_resp['Acuerdos de %s' % persona] = 'C%d' % (fila_resp + k)
        cel_resp['Acuerdos abiertos de %s' % persona] = 'D%d' % (fila_resp + k)
        cel_resp['Acuerdos vencidos de %s' % persona] = 'E%d' % (fila_resp + k)
    cel_area = {}
    for k, area in enumerate(AREAS):
        cel_area['%s: decisiones' % area] = 'C%d' % (fila_area + k)
        cel_area['%s: impacto estimado' % area] = 'D%d' % (fila_area + k)
        cel_area['%s: impacto conseguido' % area] = 'E%d' % (fila_area + k)
    cel_herr = {}
    for k, herram in enumerate(HERRAMIENTAS):
        cel_herr['%s: decisiones' % herram] = 'C%d' % (fila_herr + k)
        cel_herr['%s: impacto estimado' % herram] = 'D%d' % (fila_herr + k)
    return {
        'fichero': NOMBRE + '.xlsx',
        'hojas': {
            'Calendario de Reuniones': {
                'celdas': cel_reu,
                'tablas': [
                    {'titulo': '12 reuniones del trimestre y 12 filas libres',
                     'cols': [['#', 'A', 'num'], ['Fecha', 'B', 'txt'],
                              ['Tipo de reunión', 'C', 'txt'],
                              ['Cadencia', 'D', 'txt'],
                              ['Asistentes', 'E', 'txt'],
                              ['Duración (min)', 'F', 'num'],
                              ['Responsable', 'G', 'txt'],
                              ['Estado', 'H', 'txt'], ['Notas', 'I', 'txt']],
                     'filas': [CR0, CR1]},
                ],
            },
            'Guion de Reunión Semanal': {
                'celdas': {
                    'Puntos del guion': 'D%d' % grr,
                    'DURACIÓN TOTAL DEL GUION (min)': 'D%d' % (grr + 1),
                    'Duración objetivo de la reunión (min)': 'D%d' % (grr + 2),
                    'Diferencia con el objetivo (min)': 'D%d' % (grr + 3),
                },
                'tablas': [
                    {'titulo': 'Los 7 puntos del guion, con minutos y '
                               'responsable',
                     'cols': [['Orden', 'A', 'num'], ['Punto', 'B', 'txt'],
                              ['Minutos', 'C', 'num'],
                              ['Herramienta de la que salen los datos', 'D',
                               'txt'],
                              ['Responsable', 'E', 'txt'],
                              ['Por qué está en el guion', 'F', 'txt'],
                              ['Notas', 'G', 'txt']],
                     'filas': [GR0, GR1]},
                ],
            },
            'Uno-a-uno': {
                'celdas': {
                    'Personas del equipo en el registro': 'D%d' % urr,
                    'Sesiones con fecha anotada': 'D%d' % (urr + 1),
                    'Personas sin uno-a-uno todavía': 'D%d' % (urr + 2),
                    'EQUIPO CON UNO-A-UNO HECHO (%)': 'D%d' % (urr + 3),
                    'Objetivo de cobertura del equipo (%)': 'D%d' % (urr + 4),
                },
                'tablas': [
                    {'titulo': 'Las 6 preguntas del uno-a-uno',
                     'cols': [['#', 'A', 'num'], ['Pregunta', 'B', 'txt'],
                              ['Para qué sirve', 'C', 'txt'],
                              ['Notas', 'D', 'txt']],
                     'filas': [U_P0, U_P1]},
                    {'titulo': 'Registro de 12 sesiones, una por persona',
                     'cols': [['#', 'A', 'num'], ['Empleado', 'B', 'txt'],
                              ['Fecha', 'C', 'txt'],
                              ['Tema principal', 'D', 'txt'],
                              ['Acuerdo', 'E', 'txt'],
                              ['Fecha de seguimiento', 'F', 'txt'],
                              ['Estado', 'G', 'txt'], ['Notas', 'H', 'txt']],
                     'filas': [U_S0, U_S1]},
                ],
            },
            'Actas y Acuerdos': {
                'celdas': dict({
                    'Hoy (fecha de referencia del libro)': 'C5',
                    'Días que cuentan como «vence esta semana»': 'C6',
                    'Acuerdos registrados': 'D%d' % arr,
                    'Cerrados': 'D%d' % (arr + 1),
                    'Descartados': 'D%d' % (arr + 2),
                    'ACUERDOS ABIERTOS': 'D%d' % (arr + 3),
                    'ACUERDOS VENCIDOS': 'D%d' % (arr + 4),
                    'Vencen esta semana': 'D%d' % (arr + 5),
                    'Acuerdos cerrados (%)': 'D%d' % (arr + 6),
                    'Acuerdos en plazo (%)': 'D%d' % (arr + 7),
                    'Cerrados con fecha de cierre anotada': 'D%d' % (arr + 8),
                    'Objetivo de acuerdos en plazo (%)': 'D%d' % (arr + 12),
                }, **cel_resp),
                'tablas': [
                    {'titulo': '25 acuerdos del trimestre y 10 filas libres',
                     'cols': [['Id', 'A', 'txt'],
                              ['Fecha de la reunión', 'B', 'txt'],
                              ['Tipo de reunión', 'C', 'txt'],
                              ['Acuerdo o decisión', 'D', 'txt'],
                              ['Responsable', 'E', 'txt'],
                              ['Fecha de seguimiento', 'F', 'txt'],
                              ['Estado', 'G', 'txt'],
                              ['Fecha de cierre real', 'H', 'txt'],
                              ['Situación', 'I', 'txt'],
                              ['Días hasta el seguimiento', 'J', 'num'],
                              ['Notas', 'K', 'txt']],
                     'filas': [AA0, AA1]},
                    {'titulo': 'Acuerdos por responsable',
                     'cols': [['Responsable', 'B', 'txt'],
                              ['Acuerdos', 'C', 'num'],
                              ['Abiertos', 'D', 'num'],
                              ['Vencidos', 'E', 'num']],
                     'filas': [fila_resp, fila_resp + len(PERSONAS) - 1]},
                ],
            },
            'Plan 90 Días': {
                'celdas': dict({
                    'Decisiones registradas': 'D%d' % prr,
                    'En curso': 'D%d' % (prr + 1),
                    'Pendientes': 'D%d' % (prr + 2),
                    'Descartadas': 'D%d' % (prr + 3),
                    'Cerradas': 'D%d' % (prr + 4),
                    'Decisiones cerradas (%)': 'D%d' % (prr + 5),
                    'IMPACTO TOTAL ESTIMADO (€/mes)': 'D%d' % (prr + 6),
                    'Impacto ya conseguido (€/mes)': 'D%d' % (prr + 7),
                    'Impacto pendiente (€/mes)': 'D%d' % (prr + 8),
                    'Impacto total estimado a 12 meses (€)': 'D%d' % (prr + 9),
                    'Fecha de inicio del plan': 'D%d' % (prr + 11),
                    'Semanas del plan': 'D%d' % (prr + 12),
                    'Fecha de cierre del plan': 'D%d' % (prr + 13),
                    'Avance del calendario (%)': 'D%d' % (prr + 15),
                    'Objetivo de decisiones cerradas a 90 días (%)':
                        'D%d' % (prr + 16),
                    'Fecha objetivo de la semana 13': 'G%d' % (PL0 + 19),
                }, **dict(cel_area, **cel_herr)),
                'tablas': [
                    {'titulo': '20 decisiones del trimestre y 8 filas libres',
                     'cols': [['#', 'A', 'num'], ['Área', 'B', 'txt'],
                              ['Herramienta de origen', 'C', 'txt'],
                              ['Decisión', 'D', 'txt'],
                              ['Responsable', 'E', 'txt'],
                              ['Semana (1-13)', 'F', 'num'],
                              ['Fecha objetivo', 'G', 'txt'],
                              ['Impacto estimado (€/mes)', 'H', 'eur'],
                              ['Estado', 'I', 'txt'],
                              ['Impacto ya conseguido (€/mes)', 'J', 'eur'],
                              ['Notas', 'K', 'txt']],
                     'filas': [PL0, PL1]},
                    {'titulo': 'Resumen por área',
                     'cols': [['Área', 'B', 'txt'], ['Decisiones', 'C', 'num'],
                              ['Impacto estimado (€/mes)', 'D', 'eur'],
                              ['Impacto conseguido (€/mes)', 'E', 'eur']],
                     'filas': [fila_area, fila_area + len(AREAS) - 1]},
                    {'titulo': 'Resumen por herramienta de origen',
                     'cols': [['Herramienta de origen', 'B', 'txt'],
                              ['Decisiones', 'C', 'num'],
                              ['Impacto estimado (€/mes)', 'D', 'eur'],
                              ['Impacto conseguido (€/mes)', 'E', 'eur']],
                     'filas': [fila_herr,
                               fila_herr + len(HERRAMIENTAS) - 1]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_reuniones(wb)
    hoja_guion(wb)
    hoja_uno_a_uno(wb)
    _, fila_resp = hoja_acuerdos(wb)
    _, fila_area, fila_herr = hoja_plan(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(fila_resp, fila_area, fila_herr), fh,
                  ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-28s %d' % (hoja, n))


if __name__ == '__main__':
    main()
