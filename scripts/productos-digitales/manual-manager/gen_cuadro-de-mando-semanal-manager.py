#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_cuadro-de-mando-semanal-manager.py — libro 1 de «Manual del Manager de
Restaurante» (SPEC §2.2 fila 1; especificación celda a celda en el §3.3 del
research consolidado).

Hojas: Instrucciones · Parámetros · Semana (52 filas ISO, con gráfico) ·
KPI y Definiciones.

QUÉ HACE Y POR QUÉ NO DUPLICA AL CUADRO MENSUAL DE LA GUÍA FOOD COST
--------------------------------------------------------------------
`guia-food-cost/cuadro-de-mando-prime-cost.xlsx` mide el MES: sirve para cerrar
el ejercicio y decidir carta. Éste mide la SEMANA ISO: sirve para que el manager
vea una semana mala mientras todavía puede hacer algo. En el juego de datos de
«La Encina» el mes de septiembre cierra en el 60 % de prime cost —dentro de
objetivo— y sin embargo la semana 36 cerró en el 66,8 %: el promedio mensual se
la come. Ése es literalmente el motivo de que este libro exista, y la hoja
«Instrucciones» lo dice en voz alta para que nadie compre dos veces lo mismo.

DECISIONES TÉCNICAS
-------------------
* Objetivo ACTIVO por `INDEX`/`MATCH` sobre la lista de tipos de negocio, no
  por un `IF` binario: mañana el catálogo puede tener tres tipos y un `IF`
  anidado se rompe en silencio (SPEC §2.2, encargo del constructor).
* Los objetivos de food cost, labor cost y prime cost se ESPEJAN en la hoja
  «Semana» (bloque de PARÁMETROS EN VIGOR): el formato condicional no puede
  referirse a otra hoja en Google Sheets y la única manera sería `INDIRECT`,
  que está prohibida.
* La cotización a cargo de la empresa es una CELDA VERDE al 33 % con su
  desglose completo debajo (SPEC D5). El 23,60 % de MM-17 es SOLO contingencias
  comunes: rotularlo «coste-empresa» deja fuera casi nueve puntos y convierte
  el semáforo del prime cost en un falso negativo.
* Cero constantes dentro de una fórmula, `IFERROR(...,"")` en todo cociente,
  «sin dato» = `""` nunca 0, semáforos con `ISNUMBER`.
* Prohibidas `INDIRECT`, `COUNTA`, `PMT`, `OFFSET`, `XLOOKUP`, `LET`, `LAMBDA`
  y las matrices dinámicas: cero usos.
* Textos 100 % WinAnsi (cp1252): ni un carácter fuera, tampoco el espacio fino
  U+202F ni la flecha «→» de la nota de desproteger, que se escribe con «>».

Salida fija: build/cuadro-de-mando-semanal-manager.xlsx + su mapa de celdas.
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.normpath(os.path.join(AQUI, '..'))
sys.path.insert(0, os.path.join(RAIZ, 'guias-v2_0'))
sys.path.insert(0, AQUI)

import motor                                                   # noqa: E402
import datos_ejemplo as D                                      # noqa: E402

motor.CTX['producto'] = 'manual-manager-restaurante'

PRODUCTO = 'Manual del Manager de Restaurante'
SUBTITULO = 'AI Chef Pro · aichef.pro — ' + PRODUCTO
TITULO = 'Cuadro de mando semanal del manager'
NOMBRE = 'cuadro-de-mando-semanal-manager'

FMT_EUR = motor.FMT_EUR
FMT_PCT = motor.FMT_PCT
FMT_ENT = motor.FMT_ENT
FMT_FECHA = motor.FMT_FECHA
FMT_DEC = '#,##0.0'          # ratios (cubiertos por hora): ni € ni % ni entero

GRIS = 'F2F2F2'
CREMA = 'FFF6DC'
ORO = 'FFD700'
CABECERA = '2D2D2D'

FECHA_VERIF = '04-09-2026'


# --------------------------------------------------------------------------
def sector(idd):
    """Devuelve (dato, url, fuente) del id `MM-*` del JSON de research.

    La regla de familia es que ninguna cifra de sector se teclea en el
    generador: se lee del JSON, que lleva fuente, URL y fiabilidad.
    """
    ruta = os.path.join(RAIZ, 'auditorias', 'guias-v2-research-sector.json')
    with open(ruta, encoding='utf-8') as fh:
        datos = json.load(fh)['datos']
    for it in datos:
        if it.get('id') == idd:
            return it
    raise KeyError('id de research inexistente: ' + idd)


MM53 = sector('MM-53')
MM17 = sector('MM-17')

TIPOS_NEGOCIO = ['Sala', 'Barra o autoservicio']
#: Objetivos por tipo de negocio. Los de «Sala» son los de la casa
#: (`datos_ejemplo.RESTAURANTE`). El de labor cost de barra NO se inventa: se
#: DERIVA de los otros dos objetivos de `datos_ejemplo`
#: (prime cost de barra − food cost objetivo), de modo que las dos filas son
#: coherentes entre sí y ninguna cifra nueva entra por aquí. Los tres son
#: celdas verdes: criterio de la casa (SPEC D10), no cifra de fuente.
OBJETIVOS = {
    'Sala': (D.RESTAURANTE['food_cost_objetivo_pct'],
             D.RESTAURANTE['labor_cost_objetivo_pct'],
             D.RESTAURANTE['prime_cost_objetivo_sala']),
    'Barra o autoservicio': (
        D.RESTAURANTE['food_cost_objetivo_pct'],
        round(D.RESTAURANTE['prime_cost_objetivo_barra']
              - D.RESTAURANTE['food_cost_objetivo_pct'], 4),
        D.RESTAURANTE['prime_cost_objetivo_barra']),
}

# --- filas de la hoja «Parámetros» ---------------------------------------
P_TIPO = 5
P_CAB = 7
P_INI = 8
P_FIN = P_INI + len(TIPOS_NEGOCIO) - 1          # 9
P_FC_ACT = 11
P_LC_ACT = 12
P_PC_ACT = 13
P_SS = 22
P_DESG_CAB = 24
P_DESG_INI = 25
P_DESG_FIN = P_DESG_INI + len(D.SS_EMPRESA_DESGLOSE) - 1
P_HORAS = P_DESG_FIN + 6
P_HORAS_ANIO = P_HORAS + 1
P_HORAS_REG = P_HORAS + 2

# --- filas de la hoja «Semana» -------------------------------------------
S_CAB, S_INI = 4, 5
S_FIN = S_INI + len(D.SEMANAS) - 1              # 56
S_TOT = S_FIN + 1                               # 57
S_ESP = S_TOT + 4                               # 61 · espejo de parámetros

# --- filas de la hoja «KPI y Definiciones» -------------------------------
K_CAB, K_INI = 4, 5
K_FIN = K_INI + len(D.KPI_DEFINICIONES) - 1


# --------------------------------------------------------------------------
def cabecera(ws, fila, columnas, altura=44):
    for letra, texto in columnas:
        c = ws[letra + str(fila)]
        c.value = texto
        c.fill = PatternFill('solid', fgColor=CABECERA)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    ws.row_dimensions[fila].height = altura


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12)


def fila_total(ws, fila, primera, ultima):
    for col in range(motor.column_index_from_string(primera),
                     motor.column_index_from_string(ultima) + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = PatternFill('solid', fgColor=CREMA)
        c.font = Font(bold=True)


def cf_expresion(ws, rango, formula, bg, fg):
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


def pagina(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9                      # A4
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezar(ws, titulo, nota=None):
    motor.val(ws, 'A1', titulo)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    if nota:
        motor.val(ws, 'A3', nota)
        ws['A3'].font = Font(italic=True, size=9)


def anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


def nota(ws, fila, texto, alto=None, col='A', wrap=False):
    """Nota al pie de una sección.

    Sin `wrap` el texto se derrama sobre las celdas vacías de la derecha, que
    es como se leen las notas en las hojas estrechas (mismo criterio que el
    cuadro de mando de la Guía Food Cost). Con `wrap` hay que dar altura.
    """
    motor.val(ws, col + str(fila), texto, wrap=wrap)
    if alto:
        ws.row_dimensions[fila].height = alto


# --------------------------------------------------------------------------
PASOS = [
    '1. Hoja «Parámetros»: elige tu tipo de negocio (sala o barra y '
    'autoservicio). Los tres objetivos en vigor (food cost, labor cost y prime '
    'cost) cambian solos, y los seis de la tabla son celdas verdes que puedes '
    'ajustar a tu casa.',
    '2. Revisa la cotización a la Seguridad Social a cargo de la empresa. Viene '
    'al 33 %, que es la convención de la casa; debajo tienes el desglose '
    'partida por partida para ajustarlo a tus contratos.',
    '3. Hoja «Semana»: una fila por semana ISO. Rellena las celdas verdes con '
    'lo que te dé el TPV y la nómina: ventas de comida y de bebida, stock '
    'inicial, compras, stock final, salarios brutos, otros costes de personal, '
    'cubiertos, tickets, horas de apertura y horas trabajadas. Todo sin IVA.',
    '4. El consumo de materia prima se calcula solo: stock inicial + compras '
    'menos stock final. No son las compras de la semana; una semana con un '
    'pedido grande no es una semana cara.',
    '5. Mira la columna «Prime cost»: si se sale del objetivo, se pone en rojo '
    'y la columna de al lado lo dice con palabras. El gráfico del final enseña '
    'las 52 semanas contra la línea del objetivo.',
    '6. Las columnas de la derecha son las de servicio: ticket medio, gasto '
    'medio por cubierto, cubiertos por hora de apertura y ventas por hora '
    'trabajada. Son las que explican POR QUÉ se movió el prime cost.',
    '7. La fila TOTAL / MEDIA no hace la media de los 52 porcentajes: divide '
    'los totales del año, que es la media ponderada y la única correcta.',
    '8. Hoja «KPI y Definiciones»: qué mide cada indicador, en qué unidad, con '
    'qué cadencia y cuál es el error típico al calcularlo. Es la hoja que se '
    'consulta cuando dos personas dan dos números distintos para lo mismo.',
]

NOTAS_LIBRO = [
    'ESTE LIBRO MIDE LA SEMANA. El cuadro de mando mensual de la «Guía Food '
    'Cost + Ingeniería de Menú» (cuadro-de-mando-prime-cost.xlsx) mide el MES y '
    'sirve para cerrar el ejercicio y decidir carta. No se sustituyen: se usan '
    'en paralelo, y los dos leen el mismo prime cost.',
    'Por qué hacen falta los dos: en el restaurante de ejemplo, septiembre '
    'cierra el mes en objetivo y sin embargo la semana 36 cerró en el 66,8 % de '
    'prime cost. El promedio del mes se come la semana mala, y cuando el cierre '
    'mensual la enseña ya han pasado cuatro semanas.',
    'Todas las cifras van SIN IVA. El food cost y el prime cost se miden sobre '
    'la venta NETA: el IVA repercutido no es tuyo y el soportado se deduce en '
    'el modelo 303.',
    'Prime cost = coste de materia prima + coste de personal con Seguridad '
    'Social, sobre las ventas netas. Es la métrica que mide la salud del '
    'negocio porque junta las dos partidas que de verdad puedes mover.',
    'El margen tras prime cost no es el beneficio: de ahí todavía salen '
    'alquiler, suministros, amortizaciones e impuestos.',
    'La semana ISO empieza en lunes y la semana 1 del año es la que contiene el '
    'primer jueves de enero. Por eso la semana 1 de 2026 arranca el lunes 29 de '
    'diciembre de 2025: es la convención, no una errata.',
    'El ticket medio y el gasto por cubierto NO son lo mismo. El ticket es lo '
    'que paga una mesa; el gasto por cubierto, lo que gasta un comensal. En una '
    'mesa de dos, el ticket es el doble.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    anchos(ws, {'A': 110.0})
    motor.val(ws, 'A1', TITULO)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    motor.val(ws, 'A3', 'Para qué sirve: ver la semana mala mientras todavía '
                        'puedes hacer algo con ella.')
    ws['A3'].font = Font(italic=True, size=9)

    seccion(ws, 'A5', 'Instrucciones de uso')
    fila = 6
    for paso in PASOS:
        nota(ws, fila, paso, alto=30, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), motor.NOTA_VERDES)
    ws['A' + str(fila)].fill = PatternFill('solid', fgColor=motor.VERDE)
    fila += 2
    seccion(ws, 'A' + str(fila), 'Lo que conviene saber antes de empezar')
    fila += 1
    for texto in NOTAS_LIBRO:
        nota(ws, fila, texto, alto=44, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), D.NOTA_DESPROTEGER)
    fila += 1
    motor.val(ws, 'A' + str(fila),
              'Ruta en Excel: Revisar > Desproteger hoja. La protección no '
              'tiene contraseña; sirve para no pisar una fórmula sin querer.')
    fila += 2
    motor.val(ws, 'A' + str(fila), D.BIO)
    fila += 1
    motor.val(ws, 'A' + str(fila), D.VERSION_LINE)
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
def hoja_parametros(wb):
    ws = wb.create_sheet('Parámetros')
    anchos(ws, {'A': 46, 'B': 22, 'C': 22, 'D': 22, 'E': 58})
    encabezar(ws, 'Parámetros del cuadro de mando semanal')

    seccion(ws, 'A4', 'TIPO DE NEGOCIO Y OBJETIVOS DE LA CASA')
    motor.val(ws, 'A%d' % P_TIPO, 'Tipo de negocio', bold=True)
    motor.val(ws, 'B%d' % P_TIPO, D.RESTAURANTE['tipo_negocio'], verde_=True)
    cabecera(ws, P_CAB, [('A', 'Tipo de negocio'),
                         ('B', 'Objetivo de food cost (%)'),
                         ('C', 'Objetivo de labor cost (%)'),
                         ('D', 'Objetivo de prime cost (%)')], altura=34)
    for i, tipo in enumerate(TIPOS_NEGOCIO):
        r = P_INI + i
        fc, lc, pc = OBJETIVOS[tipo]
        motor.val(ws, 'A%d' % r, tipo, bold=True)
        motor.val(ws, 'B%d' % r, fc, fmt=FMT_PCT, verde_=True)
        motor.val(ws, 'C%d' % r, lc, fmt=FMT_PCT, verde_=True)
        motor.val(ws, 'D%d' % r, pc, fmt=FMT_PCT, verde_=True)

    for fila, col, etiqueta in ((P_FC_ACT, 'B', 'Objetivo de food cost en vigor (%)'),
                                (P_LC_ACT, 'C', 'Objetivo de labor cost en vigor (%)'),
                                (P_PC_ACT, 'D', 'Objetivo de prime cost en vigor (%)')):
        motor.val(ws, 'A%d' % fila, etiqueta, bold=True)
        motor.f(ws, 'B%d' % fila,
                '=IFERROR(INDEX(${c}${a}:${c}${b},MATCH($B${t},$A${a}:$A${b},0)),"")'
                .format(c=col, a=P_INI, b=P_FIN, t=P_TIPO), fmt=FMT_PCT)
        ws['B%d' % fila].fill = PatternFill('solid', fgColor=GRIS)

    nota(ws, 15,
         'El objetivo en vigor se busca en la tabla con INDEX y MATCH: si '
         'mañana añades un tercer tipo de negocio a la lista, basta con '
         'ampliar la tabla y la lista desplegable de arriba.')
    nota(ws, 16,
         'De dónde salen estos objetivos: la estructura de costes de '
         'referencia para la restauración española sitúa la materia prima en '
         'torno al 30 % y el personal en el 30-35 % con servicio en mesa '
         '(15-25 % en barra o autoservicio). De ahí el criterio de la casa: '
         'prime cost por debajo del 65 % con servicio en mesa y del 55 % en '
         'barra o autoservicio.')
    nota(ws, 17,
         'Ese 65 % / 55 % es CRITERIO DE LA CASA derivado de esa estructura, '
         'no una cifra publicada por nadie. Son celdas verdes: si tu convenio, '
         'tu horario o tu modelo de servicio son otros, cámbialas y el '
         'semáforo de la hoja «Semana» se recalcula.')
    nota(ws, 18,
         'Fuente de la estructura de referencia: CaixaBankLab con '
         'elBullifoundation, «Consumos y beneficios de un restaurante» (sin '
         'fecha visible en la página; revisada en 2022). '
         'https://caixabanklab.com/elbullifoundation/es/consumos-beneficios-'
         'restaurante/')

    # --- Seguridad Social a cargo de la empresa (SPEC D5) -----------------
    seccion(ws, 'A20', 'COTIZACIÓN A LA SEGURIDAD SOCIAL A CARGO DE LA EMPRESA')
    cabecera(ws, 21, [('A', 'Parámetro'), ('B', 'Valor')], altura=20)
    motor.escribir_parametro(ws, P_SS, 'A', 'B', 'ss_empresa',
                             valor=D.SS_EMPRESA)
    cabecera(ws, P_DESG_CAB, [('A', 'Partida'), ('B', 'Tipo'),
                              ('C', 'Nota')], altura=20)
    for letra in 'DE':                       # la cabecera de la nota va C:E
        c = ws[letra + str(P_DESG_CAB)]
        c.fill = PatternFill('solid', fgColor=CABECERA)
    ws.merge_cells('C%d:E%d' % (P_DESG_CAB, P_DESG_CAB))
    for i, (concepto, tipo, comentario) in enumerate(D.SS_EMPRESA_DESGLOSE):
        r = P_DESG_INI + i
        es_total = concepto.startswith('TOTAL')
        motor.val(ws, 'A%d' % r, concepto, bold=es_total, wrap=True)
        if tipo is None:
            motor.val(ws, 'B%d' % r, '')
        else:
            motor.val(ws, 'B%d' % r, tipo, fmt=FMT_PCT, bold=es_total)
        ws.merge_cells('C%d:E%d' % (r, r))
        motor.val(ws, 'C%d' % r, comentario, wrap=True)
        ws.row_dimensions[r].height = 32

    nota(ws, P_DESG_FIN + 1,
         'Verificado el ' + FECHA_VERIF + ' · ' + MM53['fuente_titulo']
         + ' · ' + MM53['url'])
    nota(ws, P_DESG_FIN + 2,
         'ATENCIÓN, es el error caro: el 23,60 % es SOLO la cotización por '
         'contingencias comunes y no debe rotularse «coste de la Seguridad '
         'Social a cargo de la empresa». El total empresarial suma además '
         'desempleo, FOGASA, formación profesional, MEI y accidentes de '
         'trabajo. Verificado el ' + FECHA_VERIF + ' · ' + MM17['fuente_titulo']
         + ' · ' + MM17['url'])

    # --- horas de apertura ------------------------------------------------
    seccion(ws, 'A%d' % (P_HORAS - 2), 'APERTURA AL PÚBLICO')
    motor.val(ws, 'A%d' % P_HORAS,
              'Horas de apertura de una semana completa (referencia)',
              bold=True)
    motor.val(ws, 'B%d' % P_HORAS,
              float(D.RESTAURANTE['horas_apertura_semana']), fmt=FMT_ENT,
              verde_=True)
    motor.val(ws, 'A%d' % P_HORAS_ANIO,
              'Horas de apertura previstas en las semanas registradas')
    motor.f(ws, 'B%d' % P_HORAS_ANIO,
            '=IFERROR(IF(OR($B${h}="",COUNT(Semana!$A${a}:$A${b})=0),"",'
            '$B${h}*COUNT(Semana!$A${a}:$A${b})),"")'
            .format(h=P_HORAS, a=S_INI, b=S_FIN), fmt=FMT_ENT)
    motor.val(ws, 'A%d' % P_HORAS_REG,
              'Horas de apertura realmente registradas en la hoja «Semana»')
    motor.f(ws, 'B%d' % P_HORAS_REG,
            '=IF(COUNT(Semana!$W${a}:$W${b})=0,"",SUM(Semana!$W${a}:$W${b}))'
            .format(a=S_INI, b=S_FIN), fmt=FMT_ENT)
    for r in (P_HORAS_ANIO, P_HORAS_REG):
        ws['B%d' % r].fill = PatternFill('solid', fgColor=GRIS)
    nota(ws, P_HORAS_REG + 1,
         'Si las horas registradas y las previstas se separan mucho es que has '
         'abierto (o cerrado) más de lo que creías: festivos, vacaciones y '
         'jornadas partidas. El dato bueno es el registrado.')

    motor.dv_lista(ws, ['B%d' % P_TIPO], TIPOS_NEGOCIO,
                   titulo='Tipo de negocio')
    motor.dv_porcentaje(ws, ['%s%d' % (c, r) for c in 'BCD'
                             for r in range(P_INI, P_FIN + 1)]
                        + ['B%d' % P_SS],
                        titulo='Porcentaje',
                        prompt='Se escribe en tanto por uno: 0,65 = 65 %.')
    motor.dv_numerica(ws, ['B%d' % P_HORAS], minimo=0, maximo=168,
                      titulo='Horas por semana')
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
COLS_SEMANA = [
    ('A', 'Semana ISO', 'ent'),
    ('B', 'Lunes de la semana', 'fecha'),
    ('C', 'Ventas netas de comida (€)', 'eur'),
    ('D', 'Ventas netas de bebida (€)', 'eur'),
    ('E', 'Ventas netas totales (€)', 'eur'),
    ('F', 'Stock inicial de materia prima (€)', 'eur'),
    ('G', 'Compras de la semana (€)', 'eur'),
    ('H', 'Stock final de materia prima (€)', 'eur'),
    ('I', 'Consumo de materia prima (€)', 'eur'),
    ('J', 'Food cost (%)', 'pct'),
    ('K', 'Salarios brutos (€)', 'eur'),
    ('L', 'Otros costes de personal (€)', 'eur'),
    ('M', 'Coste de personal con Seguridad Social (€)', 'eur'),
    ('N', 'Labor cost (%)', 'pct'),
    ('O', 'Prime cost (%)', 'pct'),
    ('P', 'Objetivo de prime cost (%)', 'pct'),
    ('Q', 'Lectura del prime cost', 'txt'),
    ('R', 'Margen tras prime cost (€)', 'eur'),
    ('S', 'Cubiertos', 'ent'),
    ('T', 'Tickets (mesas cobradas)', 'ent'),
    ('U', 'Ticket medio (€)', 'eur'),
    ('V', 'Gasto medio por cubierto (€)', 'eur'),
    ('W', 'Horas de apertura al público', 'ent'),
    ('X', 'Cubiertos por hora de apertura', 'dec'),
    ('Y', 'Horas trabajadas por la plantilla', 'ent'),
    ('Z', 'Ventas por hora trabajada (€)', 'eur'),
]

FORMATO = {'eur': FMT_EUR, 'pct': FMT_PCT, 'ent': FMT_ENT,
           'fecha': FMT_FECHA, 'dec': FMT_DEC, 'txt': None}

#: columnas de ENTRADA (celda verde) de la hoja «Semana»
VERDES = 'ABCDFGHKLSTWY'
#: columnas que se SUMAN en la fila de totales (los stocks no se suman)
SUMABLES = 'CDEGIKLMRSTWY'


def hoja_semana(wb):
    ws = wb.create_sheet('Semana')
    encabezar(ws, 'Las 52 semanas ISO del año',
              nota='Todas las cifras van sin IVA. El consumo es stock inicial '
                   '+ compras menos stock final, no las compras de la semana.')
    cabecera(ws, S_CAB, [(c, t) for c, t, _ in COLS_SEMANA], altura=70)
    anchos(ws, {'A': 9, 'B': 13, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 13,
                'H': 14, 'I': 15, 'J': 10, 'K': 13, 'L': 14, 'M': 17,
                'N': 10, 'O': 10, 'P': 12, 'Q': 22, 'R': 15, 'S': 10,
                'T': 11, 'U': 12, 'V': 13, 'W': 12, 'X': 13, 'Y': 13,
                'Z': 14})

    ss = 'Parámetros!$B$%d' % P_SS
    obj = 'Parámetros!$B$%d' % P_PC_ACT
    for i, fila in enumerate(D.SEMANAS):
        r = S_INI + i
        (sem, _mes, vcom, vbeb, ini, compras, fin, salarios, otros,
         cubiertos, tickets, h_apertura, h_trabajadas) = fila
        motor.val(ws, 'A%d' % r, sem, fmt=FMT_ENT, verde_=True)
        motor.val(ws, 'B%d' % r, D._lunes_iso(2026, sem), fmt=FMT_FECHA,
                  verde_=True)
        motor.val(ws, 'C%d' % r, float(vcom), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'D%d' % r, float(vbeb), fmt=FMT_EUR, verde_=True)
        motor.f(ws, 'E%d' % r,
                '=IFERROR(IF(AND($C{r}="",$D{r}=""),"",IF($C{r}="",0,$C{r})'
                '+IF($D{r}="",0,$D{r})),"")'.format(r=r), fmt=FMT_EUR)
        motor.val(ws, 'F%d' % r, float(ini), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'G%d' % r, float(compras), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'H%d' % r, float(fin), fmt=FMT_EUR, verde_=True)
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF(OR($F{r}="",$G{r}="",$H{r}=""),"",'
                '$F{r}+$G{r}-$H{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF(OR($I{r}="",$E{r}="",$E{r}=0),"",$I{r}/$E{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.val(ws, 'K%d' % r, float(salarios), fmt=FMT_EUR, verde_=True)
        motor.val(ws, 'L%d' % r, float(otros), fmt=FMT_EUR, verde_=True)
        motor.f(ws, 'M%d' % r,
                '=IFERROR(IF($K{r}="","",$K{r}*(1+{ss})+IF($L{r}="",0,$L{r})),'
                '"")'.format(r=r, ss=ss), fmt=FMT_EUR)
        motor.f(ws, 'N%d' % r,
                '=IFERROR(IF(OR($M{r}="",$E{r}="",$E{r}=0),"",$M{r}/$E{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'O%d' % r,
                '=IFERROR(IF(OR($I{r}="",$M{r}="",$E{r}="",$E{r}=0),"",'
                '($I{r}+$M{r})/$E{r}),"")'.format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'P%d' % r, '=IF({o}="","",{o})'.format(o=obj), fmt=FMT_PCT)
        motor.f(ws, 'Q%d' % r,
                '=IF(OR($O{r}="",$P{r}=""),"",IF($O{r}<=$P{r},"En objetivo",'
                '"Por encima del objetivo"))'.format(r=r))
        motor.f(ws, 'R%d' % r,
                '=IFERROR(IF(OR($E{r}="",$I{r}="",$M{r}=""),"",'
                '$E{r}-$I{r}-$M{r}),"")'.format(r=r), fmt=FMT_EUR)
        motor.val(ws, 'S%d' % r, cubiertos, fmt=FMT_ENT, verde_=True)
        motor.val(ws, 'T%d' % r, tickets, fmt=FMT_ENT, verde_=True)
        motor.f(ws, 'U%d' % r,
                '=IFERROR(IF(OR($E{r}="",$T{r}="",$T{r}=0),"",$E{r}/$T{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.f(ws, 'V%d' % r,
                '=IFERROR(IF(OR($E{r}="",$S{r}="",$S{r}=0),"",$E{r}/$S{r}),"")'
                .format(r=r), fmt=FMT_EUR)
        motor.val(ws, 'W%d' % r, h_apertura, fmt=FMT_ENT, verde_=True)
        motor.f(ws, 'X%d' % r,
                '=IFERROR(IF(OR($S{r}="",$W{r}="",$W{r}=0),"",$S{r}/$W{r}),"")'
                .format(r=r), fmt=FMT_DEC)
        motor.val(ws, 'Y%d' % r, h_trabajadas, fmt=FMT_ENT, verde_=True)
        motor.f(ws, 'Z%d' % r,
                '=IFERROR(IF(OR($E{r}="",$Y{r}="",$Y{r}=0),"",$E{r}/$Y{r}),"")'
                .format(r=r), fmt=FMT_EUR)

    # --- fila TOTAL / MEDIA ----------------------------------------------
    motor.val(ws, 'A%d' % S_TOT, 'TOTAL / MEDIA', bold=True)
    for col in SUMABLES:
        fmt = FMT_EUR if col in 'CDEGIKLMR' else FMT_ENT
        motor.f(ws, '%s%d' % (col, S_TOT),
                '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'
                .format(c=col, a=S_INI, b=S_FIN), fmt=fmt, bold=True)
    for col, num in (('J', 'I'), ('N', 'M')):
        motor.f(ws, '%s%d' % (col, S_TOT),
                '=IFERROR(IF(OR(${n}${t}="",$E${t}="",$E${t}=0),"",'
                '${n}${t}/$E${t}),"")'.format(n=num, t=S_TOT), fmt=FMT_PCT,
                bold=True)
    motor.f(ws, 'O%d' % S_TOT,
            '=IFERROR(IF(OR($I${t}="",$M${t}="",$E${t}="",$E${t}=0),"",'
            '($I${t}+$M${t})/$E${t}),"")'.format(t=S_TOT), fmt=FMT_PCT,
            bold=True)
    motor.f(ws, 'P%d' % S_TOT, '=IF({o}="","",{o})'.format(o=obj),
            fmt=FMT_PCT, bold=True)
    motor.f(ws, 'Q%d' % S_TOT,
            '=IF(OR($O${t}="",$P${t}=""),"",IF($O${t}<=$P${t},"En objetivo",'
            '"Por encima del objetivo"))'.format(t=S_TOT), bold=True)
    for col, num, den in (('U', 'E', 'T'), ('V', 'E', 'S'), ('X', 'S', 'W'),
                          ('Z', 'E', 'Y')):
        fmt = FMT_DEC if col == 'X' else FMT_EUR
        motor.f(ws, '%s%d' % (col, S_TOT),
                '=IFERROR(IF(OR(${n}${t}="",${d}${t}="",${d}${t}=0),"",'
                '${n}${t}/${d}${t}),"")'.format(n=num, d=den, t=S_TOT),
                fmt=fmt, bold=True)
    fila_total(ws, S_TOT, 'A', 'Z')
    nota(ws, S_TOT + 1,
         'En la fila TOTAL no se suman los stocks: el consumo del año es la '
         'suma de los consumos de las 52 semanas, y los porcentajes son medias '
         'ponderadas sobre las ventas del año, no la media de los 52 '
         'porcentajes.')

    # --- espejo de parámetros (el formato condicional no cruza de hoja) ---
    seccion(ws, 'A%d' % S_ESP,
            'PARÁMETROS EN VIGOR — se leen de la hoja «Parámetros»; cámbialos '
            'allí, no aquí')
    espejo = (('Tipo de negocio', '=Parámetros!$B$%d' % P_TIPO, None),
              ('Objetivo de food cost (%)',
               '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'
               .format(a=P_FC_ACT), FMT_PCT),
              ('Objetivo de labor cost (%)',
               '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'
               .format(a=P_LC_ACT), FMT_PCT),
              ('Objetivo de prime cost (%)',
               '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'
               .format(a=P_PC_ACT), FMT_PCT),
              ('Seguridad Social a cargo de la empresa (%)',
               '=IF(Parámetros!$B${a}="","",Parámetros!$B${a})'.format(a=P_SS),
               FMT_PCT))
    for i, (etiqueta, formula, fmt) in enumerate(espejo):
        r = S_ESP + 1 + i
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.f(ws, 'E%d' % r, formula, fmt=fmt)
        ws['E%d' % r].fill = PatternFill('solid', fgColor=GRIS)

    esp_fc = '$E$%d' % (S_ESP + 2)
    esp_lc = '$E$%d' % (S_ESP + 3)
    esp_pc = '$E$%d' % (S_ESP + 4)

    # --- semáforos --------------------------------------------------------
    for col, ancla in (('J', esp_fc), ('N', esp_lc), ('O', esp_pc)):
        rango = '{c}{a}:{c}{b}'.format(c=col, a=S_INI, b=S_TOT)
        cf_expresion(ws, rango,
                     '=AND(ISNUMBER(${c}{a}),ISNUMBER({o}),${c}{a}>{o})'
                     .format(c=col, a=S_INI, o=ancla),
                     motor.CF_ROJO_BG, motor.CF_ROJO_FG)
        cf_expresion(ws, rango,
                     '=AND(ISNUMBER(${c}{a}),ISNUMBER({o}),${c}{a}<={o})'
                     .format(c=col, a=S_INI, o=ancla),
                     motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    motor.semaforo_texto(ws, 'Q{a}:Q{b}'.format(a=S_INI, b=S_TOT),
                         (('En objetivo', motor.CF_VERDE_BG,
                           motor.CF_VERDE_FG),
                          ('Por encima del objetivo', motor.CF_ROJO_BG,
                           motor.CF_ROJO_FG)))

    # --- validaciones -----------------------------------------------------
    filas = list(range(S_INI, S_FIN + 1))
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'CDFGHKL' for r in filas],
                      minimo=0, titulo='Importe (€)')
    motor.dv_numerica(ws, ['%s%d' % (c, r) for c in 'STWY' for r in filas],
                      minimo=0, titulo='Recuento')
    motor.dv_numerica(ws, ['A%d' % r for r in filas], minimo=1, maximo=53,
                      titulo='Semana ISO',
                      mensaje='La semana ISO va de 1 a 53.')
    motor.dv_fecha(ws, ['B%d' % r for r in filas])

    # --- gráfico: prime cost vs objetivo ---------------------------------
    grafico = LineChart()
    grafico.title = 'Prime cost semanal frente al objetivo'
    grafico.style = 2
    grafico.height = 9.0
    grafico.width = 28
    datos = Reference(ws, min_col=15, max_col=16, min_row=S_CAB, max_row=S_FIN)
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(Reference(ws, min_col=1, min_row=S_INI,
                                     max_row=S_FIN))
    grafico.y_axis.numFmt = FMT_PCT
    grafico.y_axis.title = 'Prime cost (%)'
    grafico.x_axis.title = 'Semana ISO'
    grafico.x_axis.delete = False
    grafico.y_axis.delete = False
    grafico.x_axis.tickLblPos = 'low'
    grafico.y_axis.majorGridlines = None
    ws.add_chart(grafico, 'H%d' % S_ESP)

    ws.freeze_panes = 'C5'
    pagina(ws, titulos='$%d:$%d' % (S_CAB, S_CAB))
    return ws


# --------------------------------------------------------------------------
def hoja_kpi(wb):
    ws = wb.create_sheet('KPI y Definiciones')
    encabezar(ws, 'KPI y definiciones',
              nota='La columna «Error típico» es la que evita las discusiones: '
                   'casi siempre que dos personas dan dos números distintos '
                   'para lo mismo, una de las dos está cometiendo el error de '
                   'esa columna.')
    cabecera(ws, K_CAB, [('A', 'Indicador'), ('B', 'Cómo se calcula'),
                         ('C', 'Unidad'), ('D', 'Error típico'),
                         ('E', 'Cadencia')], altura=32)
    anchos(ws, {'A': 30, 'B': 56, 'C': 14, 'D': 66, 'E': 14})
    for i, (kpi, formula, unidad, error, cadencia) in enumerate(
            D.KPI_DEFINICIONES):
        r = K_INI + i
        motor.val(ws, 'A%d' % r, kpi, bold=True, wrap=True)
        motor.val(ws, 'B%d' % r, formula, wrap=True)
        motor.val(ws, 'C%d' % r, unidad, align='center')
        motor.val(ws, 'D%d' % r, error, wrap=True)
        motor.val(ws, 'E%d' % r, cadencia, align='center')
        ws.row_dimensions[r].height = 44
    nota(ws, K_FIN + 2,
         'Los ocho primeros indicadores salen de la hoja «Semana» de este mismo '
         'libro. La rotación, el absentismo y la cobertura por estación se '
         'miden con otras herramientas del pack y con otra cadencia: por eso '
         'no tienen columna aquí.')
    nota(ws, K_FIN + 3,
         'Ninguno de estos indicadores se compara con el de otro local sin '
         'mirar antes si aquel es de sala o de barra, si abre mediodía y noche '
         'y con qué convenio: comparar por comparar es la forma más rápida de '
         'tomar una decisión mala con un número bueno.')
    ws.freeze_panes = 'A5'
    pagina(ws)
    return ws


# --------------------------------------------------------------------------
def mapa():
    celdas_p = {
        'Tipo de negocio elegido': 'B%d' % P_TIPO,
        'Objetivo de food cost con servicio en mesa': 'B%d' % P_INI,
        'Objetivo de labor cost con servicio en mesa': 'C%d' % P_INI,
        'Objetivo de prime cost con servicio en mesa': 'D%d' % P_INI,
        'Objetivo de food cost en barra o autoservicio': 'B%d' % P_FIN,
        'Objetivo de labor cost en barra o autoservicio': 'C%d' % P_FIN,
        'Objetivo de prime cost en barra o autoservicio': 'D%d' % P_FIN,
        'Objetivo de food cost en vigor': 'B%d' % P_FC_ACT,
        'Objetivo de labor cost en vigor': 'B%d' % P_LC_ACT,
        'Objetivo de prime cost en vigor': 'B%d' % P_PC_ACT,
        'Seguridad Social a cargo de la empresa': 'B%d' % P_SS,
        'Contingencias comunes a cargo de la empresa': 'B%d' % P_DESG_INI,
        'Horas de apertura de una semana completa': 'B%d' % P_HORAS,
        'Horas de apertura previstas en las semanas registradas':
            'B%d' % P_HORAS_ANIO,
        'Horas de apertura registradas': 'B%d' % P_HORAS_REG,
    }
    celdas_s = {
        'Ventas netas de comida del año': 'C%d' % S_TOT,
        'Ventas netas de bebida del año': 'D%d' % S_TOT,
        'Ventas netas totales del año': 'E%d' % S_TOT,
        'Compras del año': 'G%d' % S_TOT,
        'Consumo de materia prima del año': 'I%d' % S_TOT,
        'Food cost del año (ponderado)': 'J%d' % S_TOT,
        'Salarios brutos del año': 'K%d' % S_TOT,
        'Otros costes de personal del año': 'L%d' % S_TOT,
        'Coste de personal con Seguridad Social del año': 'M%d' % S_TOT,
        'Labor cost del año (ponderado)': 'N%d' % S_TOT,
        'Prime cost del año (ponderado)': 'O%d' % S_TOT,
        'Objetivo de prime cost en vigor (espejo de la fila TOTAL)':
            'P%d' % S_TOT,
        'Lectura del prime cost del año': 'Q%d' % S_TOT,
        'Margen tras prime cost del año': 'R%d' % S_TOT,
        'Cubiertos del año': 'S%d' % S_TOT,
        'Tickets del año': 'T%d' % S_TOT,
        'Ticket medio del año': 'U%d' % S_TOT,
        'Gasto medio por cubierto del año': 'V%d' % S_TOT,
        'Horas de apertura del año': 'W%d' % S_TOT,
        'Cubiertos por hora de apertura del año': 'X%d' % S_TOT,
        'Horas trabajadas del año': 'Y%d' % S_TOT,
        'Ventas por hora trabajada del año': 'Z%d' % S_TOT,
        'Tipo de negocio (espejo)': 'E%d' % (S_ESP + 1),
        'Objetivo de food cost (espejo)': 'E%d' % (S_ESP + 2),
        'Objetivo de labor cost (espejo)': 'E%d' % (S_ESP + 3),
        'Objetivo de prime cost (espejo)': 'E%d' % (S_ESP + 4),
        'Seguridad Social a cargo de la empresa (espejo)': 'E%d' % (S_ESP + 5),
    }
    for i, filaS in enumerate(D.SEMANAS):
        sem = filaS[0]
        r = S_INI + i
        etiqueta = 'semana %d' % sem
        celdas_s['Ventas netas totales de la ' + etiqueta] = 'E%d' % r
        celdas_s['Consumo de materia prima de la ' + etiqueta] = 'I%d' % r
        celdas_s['Food cost de la ' + etiqueta] = 'J%d' % r
        celdas_s['Coste de personal con SS de la ' + etiqueta] = 'M%d' % r
        celdas_s['Labor cost de la ' + etiqueta] = 'N%d' % r
        celdas_s['Prime cost de la ' + etiqueta] = 'O%d' % r
        celdas_s['Lectura del prime cost de la ' + etiqueta] = 'Q%d' % r
        celdas_s['Margen tras prime cost de la ' + etiqueta] = 'R%d' % r
        celdas_s['Ticket medio de la ' + etiqueta] = 'U%d' % r
        celdas_s['Gasto medio por cubierto de la ' + etiqueta] = 'V%d' % r
        celdas_s['Cubiertos por hora de apertura de la ' + etiqueta] = 'X%d' % r
        celdas_s['Ventas por hora trabajada de la ' + etiqueta] = 'Z%d' % r
    return {
        'fichero': NOMBRE + '.xlsx',
        'producto': 'manual-manager-restaurante',
        'semanas_fuera_de_objetivo': list(D.SEMANAS_FUERA_DE_OBJETIVO),
        'hojas': {
            'Parámetros': {
                'celdas': celdas_p,
                'tablas': [
                    {'titulo': 'Objetivos por tipo de negocio',
                     'cols': [['Tipo de negocio', 'A', 'txt'],
                              ['Objetivo de food cost (%)', 'B', 'pct1'],
                              ['Objetivo de labor cost (%)', 'C', 'pct1'],
                              ['Objetivo de prime cost (%)', 'D', 'pct1']],
                     'filas': [P_INI, P_FIN]},
                    {'titulo': 'Desglose de la cotización a cargo de la '
                               'empresa',
                     'cols': [['Partida', 'A', 'txt'],
                              ['Tipo', 'B', 'pct1'],
                              ['Nota', 'C', 'txt']],
                     'filas': [P_DESG_INI, P_DESG_FIN]},
                ],
            },
            'Semana': {
                'celdas': celdas_s,
                'tablas': [
                    {'titulo': 'Cuadro de mando semanal',
                     'cols': [[t, c, {'eur': 'eur', 'pct': 'pct1',
                                      'ent': 'num', 'dec': 'num',
                                      'fecha': 'txt', 'txt': 'txt'}[k]]
                              for c, t, k in COLS_SEMANA],
                     'filas': [S_INI, S_TOT]},
                ],
            },
            'KPI y Definiciones': {
                'celdas': {},
                'tablas': [
                    {'titulo': 'KPI, unidad, error típico y cadencia',
                     'cols': [['Indicador', 'A', 'txt'],
                              ['Cómo se calcula', 'B', 'txt'],
                              ['Unidad', 'C', 'txt'],
                              ['Error típico', 'D', 'txt'],
                              ['Cadencia', 'E', 'txt']],
                     'filas': [K_INI, K_FIN]},
                ],
            },
        },
    }


def main():
    wb = Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_parametros(wb)
    hoja_semana(wb)
    hoja_kpi(wb)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · v1.0'

    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    destino = os.path.join(AQUI, 'build')
    os.makedirs(destino, exist_ok=True)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')
    wb.save(ruta)
    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)
    print('escrito:', ruta)
    print('formulas registradas:', len(motor.REGISTRO))


if __name__ == '__main__':
    main()
