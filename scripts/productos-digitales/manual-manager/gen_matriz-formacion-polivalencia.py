#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_matriz-formacion-polivalencia.py — libro 2 de «Manual del Manager de
Restaurante» (SPEC §2.2 fila 2; especificación celda a celda en el §3.3 del
research consolidado).

Hojas: Instrucciones · Matriz · Plan de Cross-Training · Cobertura por
Estación · Coste de una Baja.

QUÉ RESUELVE
------------
Tres preguntas que el manager contesta de memoria y casi siempre mal: quién
puede cubrir una baja sin llamar a un extra, en qué estación hay UNA SOLA
persona que sepa sostenerla, y cuánto cuesta de verdad que esa persona se vaya.
En el juego de datos de «La Encina» la respuesta de la segunda pregunta es
«Fríos y entrantes»: sólo Laura S. está a nivel 2 o más. El jefe de cocina
figura a nivel 1 y no es un descuido — la matriz registra quién puede SOSTENER
la partida en un servicio de 75 cubiertos, no quién se sabe las recetas.

DECISIONES TÉCNICAS
-------------------
* La matriz se siembra con 12 personas y 6 estaciones y admite 30 x 12: los
  nombres de las 12 estaciones son CELDAS VERDES de la fila de cabecera, y la
  hoja «Cobertura por Estación» los lee de ahí. Renombrar una estación no
  obliga a tocar ninguna fórmula.
* El nivel mínimo que da una estación por cubierta, el nivel al que se puede
  enseñar y el umbral de punto único de fallo son celdas verdes: ninguna de las
  tres cifras vive dentro de una fórmula.
* La alerta de punto único de fallo NO lleva el símbolo de aviso: «(U+26A0)»
  está fuera de WinAnsi y se pierde (o sale como un cuadrado) en cuanto el
  fichero pasa por un visor con las fuentes base. El texto lo dice con
  palabras.
* «Coste de una Baja» NO lleva ni una cifra de mercado: los nueve valores son
  del juego de ejemplo y todos editables. Separa el COSTE DIRECTO (A) del
  MARGEN que se deja de ganar (B) mientras el equipo no está a pleno
  rendimiento — no la venta bruta: A5 (auditoría 2026-09-04) corrigió que el
  bloque B aplicaba el % de caída de UN TURNO a la venta de TODO el
  restaurante, y que sumaba venta bruta con gasto real. B ya es margen tras
  prime cost, así que A + B sí es comparar euros con euros.
* `IFERROR(...,"")` en todo cociente, «sin dato» = `""` nunca 0, semáforos con
  `ISNUMBER`, cero constantes dentro de una fórmula. Prohibidas `INDIRECT`,
  `COUNTA`, `PMT`, `OFFSET`, `XLOOKUP`, `LET`, `LAMBDA` y las matrices
  dinámicas: cero usos.

Salida fija: build/matriz-formacion-polivalencia.xlsx + su mapa de celdas.
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.normpath(os.path.join(AQUI, '..'))
sys.path.insert(0, os.path.join(RAIZ, 'guias-v2_0'))
sys.path.insert(0, AQUI)

import motor                                                   # noqa: E402
import datos_ejemplo as D                                      # noqa: E402

motor.CTX['producto'] = 'manual-manager-restaurante'

PRODUCTO = 'Manual del Manager de Restaurante'
SUBTITULO = 'AI Chef Pro · aichef.pro — ' + PRODUCTO
TITULO = 'Matriz de formación y polivalencia'
NOMBRE = 'matriz-formacion-polivalencia'

FMT_EUR = motor.FMT_EUR
FMT_PCT = motor.FMT_PCT
FMT_ENT = motor.FMT_ENT
FMT_FECHA = motor.FMT_FECHA

GRIS = 'F2F2F2'
CREMA = 'FFF6DC'
ORO = 'FFD700'
CABECERA = '2D2D2D'

# --- dimensiones ----------------------------------------------------------
N_EMPLEADOS = 30                      # filas de la matriz (12 sembradas)
N_ESTACIONES = 12                     # columnas de estación (6 sembradas)
ESTADOS = ['Planificado', 'En curso', 'Completado', 'Descartado']
NIVELES = [str(n) for n, _ in D.NIVELES_POLIVALENCIA]

#: nivel al que una persona puede SOSTENER la estación en un servicio
NIVEL_CUBRE = 2
#: nivel al que una persona puede ENSEÑAR la estación
NIVEL_ENSENA = max(n for n, _ in D.NIVELES_POLIVALENCIA)      # 3
#: por debajo de cuántas personas una estación es un punto único de fallo
UMBRAL_RIESGO = 1
#: antelación con la que el plan avisa de una fecha objetivo (días, editable)
AVISO_DIAS = 30

# --- filas de la hoja «Matriz» -------------------------------------------
M_LEY_CAB = 5
M_LEY_INI = 6
M_LEY_FIN = M_LEY_INI + len(D.NIVELES_POLIVALENCIA) - 1        # 9
M_PAR_CAB = 11
M_CUBRE = 12
M_ENSENA = 13
M_RIESGO = 14
M_CAB = 17
M_INI = 18
M_FIN = M_INI + N_EMPLEADOS - 1                                # 47
COL_EST0 = 5                                                   # columna E
COL_EST1 = COL_EST0 + N_ESTACIONES - 1                         # columna P
COL_SOST = get_column_letter(COL_EST1 + 1)                     # Q
COL_ENS = get_column_letter(COL_EST1 + 2)                      # R

# --- filas de la hoja «Plan de Cross-Training» ---------------------------
X_CORTE = 5
X_AVISO = 6
X_CAB = 9
X_INI = 10
N_PLAN = 20
X_FIN = X_INI + N_PLAN - 1                                     # 29
X_RES = X_FIN + 2                                              # 31

# --- filas de la hoja «Cobertura por Estación» ---------------------------
C_ESP = 5
C_CUBRE = C_ESP + 1
C_ENSENA = C_ESP + 2
C_RIESGO = C_ESP + 3
C_CAB = 11
C_INI = 12
C_FIN = C_INI + N_ESTACIONES - 1                               # 23
C_RES = C_FIN + 2                                              # 25

# --- filas de la hoja «Coste de una Baja» --------------------------------
B_A_CAB = 5
B_H_SEL = 6
B_E_SEL = 7
B_T_SEL = 8
B_H_FOR = 9
B_E_FOR = 10
B_T_FOR = 11
B_H_NUE = 12
B_E_NUE = 13
B_T_NUE = 14
B_TOT_A = 15
B_B_CAB = 18
B_DIAS = 19
B_PCT = 20
B_VENTA = 21
B_TOT_B = 22
B_C_CAB = 25
B_REF_A = 26
B_REF_B = 27
B_TOTAL = 28


# --------------------------------------------------------------------------
def cabecera(ws, fila, columnas, altura=44):
    for letra, texto in columnas:
        c = ws[letra + str(fila)]
        c.value = texto
        c.fill = PatternFill('solid', fgColor=CABECERA)
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    ws.row_dimensions[fila].height = altura


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12)


def fila_total(ws, fila, primera, ultima):
    for col in range(motor.column_index_from_string(primera),
                     motor.column_index_from_string(ultima) + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = PatternFill('solid', fgColor=CREMA)
        c.font = Font(bold=True)


def cf_expresion(ws, rango, formula, bg, fg):
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


def pagina(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9                      # A4
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezar(ws, titulo, nota_=None):
    motor.val(ws, 'A1', titulo)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    if nota_:
        motor.val(ws, 'A3', nota_)
        ws['A3'].font = Font(italic=True, size=9)


def anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


def nota(ws, fila, texto, alto=None, col='A', wrap=False):
    motor.val(ws, col + str(fila), texto, wrap=wrap)
    if alto:
        ws.row_dimensions[fila].height = alto


# --------------------------------------------------------------------------
PASOS = [
    '1. Hoja «Matriz»: escribe a tu gente en las filas y tus estaciones en la '
    'fila verde de cabecera. Vienen sembradas 12 personas y 6 estaciones; caben '
    '30 personas y 12 estaciones sin tocar nada.',
    '2. Puntúa cada cruce de 0 a 3 con el desplegable. La leyenda está arriba: '
    '0 no formado, 1 formado pero necesita supervisión, 2 autónomo, 3 puede '
    'formar a otra persona.',
    '3. Puntúa quién SOSTIENE la estación en un servicio lleno, no quién ha '
    'estado alguna vez. Es el error que vacía de sentido toda la matriz: si '
    'todos salen a 2, la matriz no avisa de nada.',
    '4. Hoja «Cobertura por Estación»: te dice cuántas personas sostienen cada '
    'estación y te enciende la alerta cuando sólo hay una. Ése es tu riesgo '
    'real: el día que esa persona se pone mala, cierras esa partida.',
    '5. Hoja «Plan de Cross-Training»: una fila por acción de formación. Quién, '
    'qué estación, de qué nivel a qué nivel, quién le enseña y para cuándo. El '
    'libro calcula los días que faltan y avisa de lo vencido.',
    '6. Hoja «Coste de una Baja»: pon TUS horas y TUS costes por hora. Ninguna '
    'de las cifras que trae es una referencia de mercado; son un ejemplo para '
    'que veas el orden de magnitud y lo rehagas con tus números.',
    '7. Revisa la matriz una vez al trimestre, no cuando ya tienes el problema '
    'encima. Una estación pasa de cubierta a descubierta con una sola baja.',
]

NOTAS_LIBRO = [
    'Qué mide un nivel 2: que esa persona puede llevar la estación ELLA SOLA en '
    'un servicio normal, sin que nadie la esté corrigiendo. Un nivel 1 no cubre '
    'un servicio, y contarlo como si lo cubriera es exactamente lo que hace que '
    'la alerta de esta hoja no salte cuando debería.',
    'Punto único de fallo: una estación que sólo sostiene una persona. No es un '
    'problema de esa persona, es un problema de organización, y se arregla '
    'formando a alguien más con el plan de cross-training de este mismo libro.',
    'El nivel mínimo para dar una estación por cubierta, el nivel al que se '
    'puede enseñar y el umbral de riesgo son celdas verdes de la hoja «Matriz». '
    'Si en tu casa hacen falta dos personas por estación, cámbialo ahí y las '
    'alertas se recalculan.',
    'Los nombres de las estaciones son celdas verdes de la fila de cabecera de '
    'la matriz. Cámbialos por los tuyos: la hoja de cobertura los lee de ahí, '
    'no hay que tocar ninguna fórmula.',
    # A5 (auditoría 2026-09-04): esta nota describía el bloque B como VENTA
    # bruta sumada a un GASTO —justo el error que corrigió A5—. Ahora que B
    # es MARGEN tras prime cost, las dos partidas SÍ son la misma clase de
    # euro y sumarlas es correcto; la nota lo dice.
    'La hoja «Coste de una Baja» tiene dos partidas que juntas responden a '
    '«cuánto le cuesta al negocio perder a una persona»: el COSTE DIRECTO '
    '(horas de selección y de formación) y el MARGEN que se deja de ganar '
    'mientras el equipo no está a pleno rendimiento. Las dos son euros reales '
    '—uno sale de la cuenta, el otro deja de entrar—, así que sumarlas en el '
    'total sí es comparar euros con euros.',
    'Esta matriz no evalúa a nadie ni sirve para decidir un despido: mide qué '
    'sabe hacer el equipo y dónde hay que formar. La evaluación del desempeño '
    'es otra herramienta y otra conversación.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    anchos(ws, {'A': 110.0})
    motor.val(ws, 'A1', TITULO)
    ws['A1'].font = Font(bold=True, size=16, color=ORO)
    ws.row_dimensions[1].height = 30
    motor.val(ws, 'A2', SUBTITULO)
    motor.val(ws, 'A3', 'Para qué sirve: saber quién cubre una baja, dónde hay '
                        'una sola persona que sepa hacer algo y cuánto cuesta '
                        'perderla.')
    ws['A3'].font = Font(italic=True, size=9)

    seccion(ws, 'A5', 'Instrucciones de uso')
    fila = 6
    for paso in PASOS:
        nota(ws, fila, paso, alto=30, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), motor.NOTA_VERDES)
    ws['A' + str(fila)].fill = PatternFill('solid', fgColor=motor.VERDE)
    fila += 2
    seccion(ws, 'A' + str(fila), 'Lo que conviene saber antes de empezar')
    fila += 1
    for texto in NOTAS_LIBRO:
        nota(ws, fila, texto, alto=44, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), D.NOTA_DESPROTEGER)
    fila += 1
    motor.val(ws, 'A' + str(fila),
              'Ruta en Excel: Revisar > Desproteger hoja. La protección no '
              'tiene contraseña; sirve para no pisar una fórmula sin querer.')
    fila += 2
    motor.val(ws, 'A' + str(fila), D.BIO)
    fila += 1
    motor.val(ws, 'A' + str(fila), D.VERSION_LINE)
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
def hoja_matriz(wb):
    ws = wb.create_sheet('Matriz')
    encabezar(ws, 'Matriz de polivalencia',
              nota_='Puntúa quién puede SOSTENER la estación en un servicio '
                    'lleno, no quién ha estado alguna vez.')
    anchos(ws, {'A': 8, 'B': 20, 'C': 34, 'D': 22})
    for i in range(N_ESTACIONES):
        ws.column_dimensions[get_column_letter(COL_EST0 + i)].width = 13
    ws.column_dimensions[COL_SOST].width = 14
    ws.column_dimensions[COL_ENS].width = 14

    # --- leyenda ----------------------------------------------------------
    cabecera(ws, M_LEY_CAB, [('A', 'Nivel'), ('B', 'Qué significa')],
             altura=20)
    for letra in ('C', 'D'):
        ws[letra + str(M_LEY_CAB)].fill = PatternFill('solid',
                                                      fgColor=CABECERA)
    ws.merge_cells('B%d:D%d' % (M_LEY_CAB, M_LEY_CAB))
    for i, (nivel, texto) in enumerate(D.NIVELES_POLIVALENCIA):
        r = M_LEY_INI + i
        motor.val(ws, 'A%d' % r, nivel, fmt=FMT_ENT, align='center', bold=True)
        ws.merge_cells('B%d:D%d' % (r, r))
        motor.val(ws, 'B%d' % r, texto)

    # --- parámetros -------------------------------------------------------
    cabecera(ws, M_PAR_CAB, [('A', 'Parámetro'), ('B', 'Valor')], altura=20)
    for letra in ('C', 'D'):
        ws[letra + str(M_PAR_CAB)].fill = PatternFill('solid',
                                                      fgColor=CABECERA)
    ws.merge_cells('C%d:D%d' % (M_PAR_CAB, M_PAR_CAB))
    params = ((M_CUBRE, 'Nivel mínimo para dar una estación por cubierta',
               NIVEL_CUBRE),
              (M_ENSENA, 'Nivel a partir del cual se puede enseñar la estación',
               NIVEL_ENSENA),
              (M_RIESGO, 'Punto único de fallo: personas o menos por estación',
               UMBRAL_RIESGO))
    for r, etiqueta, valor in params:
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.val(ws, 'B%d' % r, valor, fmt=FMT_ENT, verde_=True)
    nota(ws, M_RIESGO + 1,
         'Los tres son celdas verdes: ninguna de estas cifras vive dentro de '
         'una fórmula. Si en tu casa una estación no está cubierta hasta que '
         'hay dos personas, sube el umbral de riesgo a 2.')

    # --- cabecera de la matriz -------------------------------------------
    fijas = [('A', 'Id'), ('B', 'Nombre'), ('C', 'Puesto'),
             ('D', 'Estación principal')]
    cabecera(ws, M_CAB, fijas + [(COL_SOST, 'Estaciones que puede sostener'),
                                 (COL_ENS, 'Estaciones que puede enseñar')],
             altura=56)
    for i in range(N_ESTACIONES):
        letra = get_column_letter(COL_EST0 + i)
        valor = D.ESTACIONES[i] if i < len(D.ESTACIONES) else None
        c = motor.val(ws, letra + str(M_CAB), valor, verde_=True, bold=True,
                      align='center', wrap=True)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)

    # --- filas de personas ------------------------------------------------
    nivel_por_id = dict(D.POLIVALENCIA)
    for i in range(N_EMPLEADOS):
        r = M_INI + i
        if i < len(D.PLANTILLA):
            (pid, nombre, puesto, _area, _grupo, _contrato, _jornada,
             _bruto, _alta, estacion) = D.PLANTILLA[i]
            motor.val(ws, 'A%d' % r, pid, verde_=True, align='center')
            motor.val(ws, 'B%d' % r, nombre, verde_=True)
            motor.val(ws, 'C%d' % r, puesto, verde_=True)
            motor.val(ws, 'D%d' % r, estacion, verde_=True)
            niveles = nivel_por_id[pid]
        else:
            for letra in 'ABCD':
                motor.verde(ws, '%s%d' % (letra, r))
            niveles = []
        for j in range(N_ESTACIONES):
            letra = get_column_letter(COL_EST0 + j)
            if j < len(niveles):
                motor.val(ws, letra + str(r), niveles[j], fmt=FMT_ENT,
                          verde_=True, align='center')
            else:
                motor.verde(ws, letra + str(r))
                ws[letra + str(r)].number_format = FMT_ENT
                ws[letra + str(r)].alignment = Alignment(horizontal='center')
        motor.f(ws, '%s%d' % (COL_SOST, r),
                '=IF(OR($A{r}="",ISNUMBER($B${p})=FALSE),"",'
                'COUNTIFS(${e}{r}:${f}{r},">="&$B${p}))'
                .format(r=r, p=M_CUBRE, e=get_column_letter(COL_EST0),
                        f=get_column_letter(COL_EST1)), fmt=FMT_ENT)
        motor.f(ws, '%s%d' % (COL_ENS, r),
                '=IF(OR($A{r}="",ISNUMBER($B${p})=FALSE),"",'
                'COUNTIFS(${e}{r}:${f}{r},">="&$B${p}))'
                .format(r=r, p=M_ENSENA, e=get_column_letter(COL_EST0),
                        f=get_column_letter(COL_EST1)), fmt=FMT_ENT)

    # --- semáforo de niveles ---------------------------------------------
    rango = '{e}{a}:{f}{b}'.format(e=get_column_letter(COL_EST0), a=M_INI,
                                   f=get_column_letter(COL_EST1), b=M_FIN)
    ancla = '%s%d' % (get_column_letter(COL_EST0), M_INI)
    cf_expresion(ws, rango,
                 '=AND(ISNUMBER({c}),ISNUMBER($B${p}),{c}>=$B${p})'
                 .format(c=ancla, p=M_CUBRE),
                 motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    cf_expresion(ws, rango,
                 '=AND(ISNUMBER({c}),ISNUMBER($B${p}),{c}>0,{c}<$B${p})'
                 .format(c=ancla, p=M_CUBRE),
                 motor.CF_AMBAR_BG, motor.CF_AMBAR_FG)

    # --- validaciones -----------------------------------------------------
    filas = list(range(M_INI, M_FIN + 1))
    motor.dv_lista(ws, ['%s%d' % (get_column_letter(COL_EST0 + j), r)
                        for j in range(N_ESTACIONES) for r in filas],
                   NIVELES, titulo='Nivel de 0 a 3',
                   mensaje='El nivel va de 0 a 3. Mira la leyenda de arriba.')
    # A4 (auditoría 2026-09-04): `B14` (M_RIESGO) es un RECUENTO de personas
    # (su propia DV decimal va 0-30 líneas más abajo), no un nivel 0-3. Antes
    # se colaba también en el `sqref` de esta lista, y dos `dataValidation`
    # solapadas sobre la misma celda viola ECMA-376 §18.3.1.32. Sólo
    # M_CUBRE y M_ENSENA (los dos primeros de `params`) son niveles.
    motor.dv_lista(ws, ['B%d' % r for r, _, _ in params if r != M_RIESGO],
                   NIVELES, titulo='Nivel de 0 a 3')
    motor.dv_numerica(ws, ['B%d' % M_RIESGO], minimo=0, maximo=N_EMPLEADOS,
                      titulo='Personas')
    ws.freeze_panes = 'E%d' % M_INI
    pagina(ws, titulos='$%d:$%d' % (M_CAB, M_CAB))
    return ws


# --------------------------------------------------------------------------
def hoja_plan(wb):
    ws = wb.create_sheet('Plan de Cross-Training')
    encabezar(ws, 'Plan de cross-training',
              nota_='Una fila por acción de formación: quién, qué estación, de '
                    'qué nivel a qué nivel, quién le enseña y para cuándo.')
    anchos(ws, {'A': 6, 'B': 12, 'C': 20, 'D': 24, 'E': 12, 'F': 12,
                'G': 22, 'H': 14, 'I': 14, 'J': 13, 'K': 14, 'L': 18})

    motor.val(ws, 'A%d' % X_CORTE, 'Fecha de corte (desde cuándo se cuentan '
                                   'los días)')
    motor.val(ws, 'C%d' % X_CORTE, D._fecha(D.RESTAURANTE['fecha_corte_normativa']),
              fmt=FMT_FECHA, verde_=True)
    motor.val(ws, 'A%d' % X_AVISO, 'Avisar cuando falten estos días o menos')
    motor.val(ws, 'C%d' % X_AVISO, AVISO_DIAS, fmt=FMT_ENT, verde_=True)
    nota(ws, X_AVISO + 1,
         'Las dos son celdas verdes. La fecha de corte no se toma del reloj a '
         'propósito: así el libro enseña siempre lo mismo cada vez que se abre '
         'y no cambia de color solo.')

    cabecera(ws, X_CAB, [('A', '#'), ('B', 'Id del empleado'), ('C', 'Nombre'),
                         ('D', 'Estación objetivo'), ('E', 'Nivel actual'),
                         ('F', 'Nivel objetivo'),
                         ('G', 'Quién le enseña'), ('H', 'Fecha objetivo'),
                         ('I', 'Estado'), ('J', 'Niveles que faltan'),
                         ('K', 'Días hasta la fecha'),
                         ('L', 'Situación del plazo')], altura=56)
    nombre_por_id = {p[0]: p[1] for p in D.PLANTILLA}
    for i in range(N_PLAN):
        r = X_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=FMT_ENT, align='center')
        if i < len(D.PLAN_CROSS_TRAINING):
            (pid, estacion, actual, objetivo, responsable, fecha,
             estado) = D.PLAN_CROSS_TRAINING[i]
            motor.val(ws, 'B%d' % r, pid, verde_=True, align='center')
            motor.val(ws, 'C%d' % r, nombre_por_id[pid], verde_=True)
            motor.val(ws, 'D%d' % r, estacion, verde_=True)
            motor.val(ws, 'E%d' % r, actual, fmt=FMT_ENT, verde_=True,
                      align='center')
            motor.val(ws, 'F%d' % r, objetivo, fmt=FMT_ENT, verde_=True,
                      align='center')
            motor.val(ws, 'G%d' % r, nombre_por_id[responsable], verde_=True)
            motor.val(ws, 'H%d' % r, D._fecha(fecha), fmt=FMT_FECHA,
                      verde_=True)
            motor.val(ws, 'I%d' % r, estado, verde_=True, align='center')
        else:
            for letra in 'BCDG':
                motor.verde(ws, '%s%d' % (letra, r))
            for letra in 'EF':
                motor.verde(ws, '%s%d' % (letra, r))
                ws['%s%d' % (letra, r)].number_format = FMT_ENT
            motor.verde(ws, 'H%d' % r)
            ws['H%d' % r].number_format = FMT_FECHA
            motor.verde(ws, 'I%d' % r)
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF(OR($E{r}="",$F{r}=""),"",$F{r}-$E{r}),"")'
                .format(r=r), fmt=FMT_ENT)
        motor.f(ws, 'K%d' % r,
                '=IFERROR(IF(OR($H{r}="",$C${c}=""),"",$H{r}-$C${c}),"")'
                .format(r=r, c=X_CORTE), fmt=FMT_ENT)
        motor.f(ws, 'L%d' % r,
                '=IF($K{r}="","",IF($I{r}="Completado","Completado",'
                'IF($I{r}="Descartado","Descartado",'
                'IF($K{r}<0,"Vencido",'
                'IF(AND(ISNUMBER($C${a}),$K{r}<=$C${a}),"Vence pronto",'
                '"En plazo")))))'.format(r=r, a=X_AVISO))

    # --- resumen ----------------------------------------------------------
    seccion(ws, 'A%d' % X_RES, 'RESUMEN DEL PLAN')
    resumen = (
        ('Acciones registradas',
         '=COUNTIF($B${a}:$B${b},"<>")', FMT_ENT),
        ('Completadas', '=COUNTIF($I${a}:$I${b},"Completado")', FMT_ENT),
        ('En curso', '=COUNTIF($I${a}:$I${b},"En curso")', FMT_ENT),
        ('Planificadas', '=COUNTIF($I${a}:$I${b},"Planificado")', FMT_ENT),
        ('Vencidas', '=COUNTIF($L${a}:$L${b},"Vencido")', FMT_ENT),
        ('Que vencen pronto', '=COUNTIF($L${a}:$L${b},"Vence pronto")',
         FMT_ENT),
    )
    for i, (etiqueta, formula, fmt) in enumerate(resumen):
        r = X_RES + 1 + i
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.f(ws, 'C%d' % r, formula.format(a=X_INI, b=X_FIN), fmt=fmt)
        ws['C%d' % r].fill = PatternFill('solid', fgColor=GRIS)
    r_pct = X_RES + 1 + len(resumen)
    motor.val(ws, 'A%d' % r_pct, 'Acciones completadas (%)')
    motor.f(ws, 'C%d' % r_pct,
            '=IFERROR(IF(OR($C${t}="",$C${t}=0),"",$C${c}/$C${t}),"")'
            .format(t=X_RES + 1, c=X_RES + 2), fmt=FMT_PCT)
    ws['C%d' % r_pct].fill = PatternFill('solid', fgColor=GRIS)

    # --- semáforos --------------------------------------------------------
    motor.semaforo_texto(ws, 'L{a}:L{b}'.format(a=X_INI, b=X_FIN),
                         (('Vencido', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
                          ('Vence pronto', motor.CF_AMBAR_BG,
                           motor.CF_AMBAR_FG),
                          ('En plazo', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
                          ('Completado', motor.CF_VERDE_BG,
                           motor.CF_VERDE_FG),
                          ('Descartado', motor.CF_GRIS_BG,
                           motor.CF_GRIS_FG)))
    motor.semaforo_texto(ws, 'I{a}:I{b}'.format(a=X_INI, b=X_FIN),
                         (('Completado', motor.CF_VERDE_BG,
                           motor.CF_VERDE_FG),
                          ('En curso', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
                          ('Descartado', motor.CF_GRIS_BG,
                           motor.CF_GRIS_FG)))

    # --- validaciones -----------------------------------------------------
    filas = list(range(X_INI, X_FIN + 1))
    motor.dv_lista(ws, ['%s%d' % (c, r) for c in 'EF' for r in filas],
                   NIVELES, titulo='Nivel de 0 a 3')
    motor.dv_lista(ws, ['I%d' % r for r in filas], ESTADOS,
                   titulo='Estado de la acción')
    motor.dv_fecha(ws, ['H%d' % r for r in filas] + ['C%d' % X_CORTE])
    motor.dv_numerica(ws, ['C%d' % X_AVISO], minimo=0, maximo=365,
                      titulo='Días de aviso')
    ws.freeze_panes = 'B%d' % X_INI
    pagina(ws, titulos='$%d:$%d' % (X_CAB, X_CAB))
    return ws


# --------------------------------------------------------------------------
def hoja_cobertura(wb):
    ws = wb.create_sheet('Cobertura por Estación')
    encabezar(ws, 'Cobertura por estación',
              nota_='Se calcula sola desde la hoja «Matriz». Si una estación '
                    'tiene una sola persona que la sostenga, aquí se enciende.')
    anchos(ws, {'A': 26, 'B': 16, 'C': 16, 'D': 16, 'E': 14, 'F': 62})

    seccion(ws, 'A%d' % C_ESP,
            'PARÁMETROS EN VIGOR — se leen de la hoja «Matriz»; cámbialos '
            'allí, no aquí')
    espejo = ((C_CUBRE, 'Nivel mínimo para dar una estación por cubierta',
               M_CUBRE),
              (C_ENSENA, 'Nivel a partir del cual se puede enseñar', M_ENSENA),
              (C_RIESGO, 'Punto único de fallo: personas o menos', M_RIESGO))
    for r, etiqueta, origen in espejo:
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.f(ws, 'C%d' % r,
                '=IF(Matriz!$B${o}="","",Matriz!$B${o})'.format(o=origen),
                fmt=FMT_ENT)
        ws['C%d' % r].fill = PatternFill('solid', fgColor=GRIS)
    nota(ws, C_RIESGO + 1,
         'Se copian aquí porque el formato condicional no puede leer otra '
         'hoja: los colores de esta tabla necesitan tener el umbral al lado.')

    cabecera(ws, C_CAB, [('A', 'Estación'),
                         ('B', 'Personas que la sostienen'),
                         ('C', 'Personas que pueden enseñarla'),
                         ('D', 'Personas en la matriz'),
                         ('E', 'Cobertura (%)'),
                         ('F', 'Alerta')], altura=56)
    for i in range(N_ESTACIONES):
        r = C_INI + i
        col = get_column_letter(COL_EST0 + i)
        motor.f(ws, 'A%d' % r,
                '=IF(Matriz!{c}${h}="","",Matriz!{c}${h})'
                .format(c=col, h=M_CAB))
        motor.f(ws, 'B%d' % r,
                '=IF(OR($A{r}="",ISNUMBER($C${p})=FALSE),"",'
                'COUNTIFS(Matriz!${c}${a}:${c}${b},">="&$C${p}))'
                .format(r=r, p=C_CUBRE, c=col, a=M_INI, b=M_FIN), fmt=FMT_ENT)
        motor.f(ws, 'C%d' % r,
                '=IF(OR($A{r}="",ISNUMBER($C${p})=FALSE),"",'
                'COUNTIFS(Matriz!${c}${a}:${c}${b},">="&$C${p}))'
                .format(r=r, p=C_ENSENA, c=col, a=M_INI, b=M_FIN),
                fmt=FMT_ENT)
        motor.f(ws, 'D%d' % r,
                '=IF($A{r}="","",COUNTIF(Matriz!$A${a}:$A${b},"<>"))'
                .format(r=r, a=M_INI, b=M_FIN), fmt=FMT_ENT)
        motor.f(ws, 'E%d' % r,
                '=IFERROR(IF(OR($B{r}="",$D{r}="",$D{r}=0),"",$B{r}/$D{r}),"")'
                .format(r=r), fmt=FMT_PCT)
        motor.f(ws, 'F%d' % r,
                '=IF($B{r}="","",'
                'IF($B{r}=0,"RIESGO: nadie puede sostener esta estación",'
                'IF($B{r}=1,"RIESGO: punto único de fallo, solo 1 persona '
                'puede sostener esta estación",'
                'IF(AND(ISNUMBER($C${u}),$B{r}<=$C${u}),'
                '"Cobertura por debajo del umbral","Cobertura suficiente"))))'
                .format(r=r, u=C_RIESGO))

    # --- resumen ----------------------------------------------------------
    seccion(ws, 'A%d' % C_RES, 'RESUMEN')
    resumen = (
        ('Estaciones registradas', '=COUNTIF($A${a}:$A${b},"<>")', FMT_ENT),
        ('Personas registradas en la matriz',
         '=COUNTIF(Matriz!$A${m}:$A${n},"<>")', FMT_ENT),
        ('Estaciones en riesgo (en el umbral o por debajo)',
         '=IF(ISNUMBER($C${u})=FALSE,"",SUMPRODUCT(--(ISNUMBER($B${a}:$B${b})),'
         '--($B${a}:$B${b}<=$C${u})))', FMT_ENT),
        ('Personas que sostienen la estación mejor cubierta',
         '=IF(COUNT($B${a}:$B${b})=0,"",MAX($B${a}:$B${b}))', FMT_ENT),
        ('Personas que sostienen la estación peor cubierta',
         '=IF(COUNT($B${a}:$B${b})=0,"",MIN($B${a}:$B${b}))', FMT_ENT),
    )
    for i, (etiqueta, formula, fmt) in enumerate(resumen):
        r = C_RES + 1 + i
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.f(ws, 'C%d' % r,
                formula.format(a=C_INI, b=C_FIN, u=C_RIESGO, m=M_INI, n=M_FIN),
                fmt=fmt)
        ws['C%d' % r].fill = PatternFill('solid', fgColor=GRIS)
    r_pct = C_RES + 1 + len(resumen)
    motor.val(ws, 'A%d' % r_pct, 'Estaciones en riesgo (%)')
    motor.f(ws, 'C%d' % r_pct,
            '=IFERROR(IF(OR($C${t}="",$C${t}=0,$C${g}=""),"",'
            '$C${g}/$C${t}),"")'.format(t=C_RES + 1, g=C_RES + 3),
            fmt=FMT_PCT)
    ws['C%d' % r_pct].fill = PatternFill('solid', fgColor=GRIS)
    nota(ws, r_pct + 2,
         'Una estación en riesgo no se arregla contratando: se arregla '
         'formando. Llévala a la hoja «Plan de Cross-Training» con nombre, '
         'responsable y fecha.')

    # --- semáforos --------------------------------------------------------
    motor.semaforo_texto(
        ws, 'F{a}:F{b}'.format(a=C_INI, b=C_FIN),
        (('RIESGO: nadie puede sostener esta estación', motor.CF_ROJO_BG,
          motor.CF_ROJO_FG),
         ('RIESGO: punto único de fallo, solo 1 persona puede sostener esta '
          'estación', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
         ('Cobertura por debajo del umbral', motor.CF_AMBAR_BG,
          motor.CF_AMBAR_FG),
         ('Cobertura suficiente', motor.CF_VERDE_BG, motor.CF_VERDE_FG)))
    cf_expresion(ws, 'B{a}:B{b}'.format(a=C_INI, b=C_FIN),
                 '=AND(ISNUMBER($B{a}),ISNUMBER($C${u}),$B{a}<=$C${u})'
                 .format(a=C_INI, u=C_RIESGO),
                 motor.CF_ROJO_BG, motor.CF_ROJO_FG)
    cf_expresion(ws, 'B{a}:B{b}'.format(a=C_INI, b=C_FIN),
                 '=AND(ISNUMBER($B{a}),ISNUMBER($C${u}),$B{a}>$C${u})'
                 .format(a=C_INI, u=C_RIESGO),
                 motor.CF_VERDE_BG, motor.CF_VERDE_FG)
    ws.freeze_panes = 'B%d' % C_INI
    pagina(ws, titulos='$%d:$%d' % (C_CAB, C_CAB))
    return ws


# --------------------------------------------------------------------------
def hoja_coste_baja(wb):
    ws = wb.create_sheet('Coste de una Baja')
    encabezar(ws, 'Coste de una baja',
              nota_='Todas las cifras son de ejemplo y todas son editables: '
                    'aquí no hay ni una referencia de mercado. Pon las tuyas.')
    anchos(ws, {'A': 54, 'B': 18, 'C': 66})

    cb = D.COSTE_BAJA
    seccion(ws, 'A%d' % (B_A_CAB - 1),
            'A. COSTE DIRECTO DE CUBRIR LA BAJA')
    cabecera(ws, B_A_CAB, [('A', 'Concepto'), ('B', 'Valor'), ('C', 'Nota')],
             altura=20)
    directos = (
        (B_H_SEL, 'Horas dedicadas a seleccionar y contratar',
         float(cb['horas_seleccion']), FMT_ENT,
         'Publicar la oferta, cribar, entrevistar y hacer el alta.'),
        (B_E_SEL, 'Coste por hora de quien selecciona',
         cb['coste_hora_seleccion'], FMT_EUR,
         'Coste-empresa por hora de quien dedica ese tiempo, no su salario '
         'bruto por hora.'),
        (B_H_FOR, 'Horas que dedica quien enseña',
         float(cb['horas_formacion_formador']), FMT_ENT,
         'Horas en las que el formador está enseñando en vez de produciendo.'),
        (B_E_FOR, 'Coste por hora de quien enseña', cb['coste_hora_formador'],
         FMT_EUR, ''),
        (B_H_NUE, 'Horas de formación de la persona nueva',
         float(cb['horas_formacion_formado']), FMT_ENT,
         'Horas pagadas en las que todavía no rinde como el puesto pide.'),
        (B_E_NUE, 'Coste por hora de la persona nueva',
         cb['coste_hora_formado'], FMT_EUR, ''),
    )
    for r, etiqueta, valor, fmt, comentario in directos:
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.val(ws, 'B%d' % r, valor, fmt=fmt, verde_=True)
        if comentario:
            motor.val(ws, 'C%d' % r, comentario, wrap=True)
            ws.row_dimensions[r].height = 30
    for r, etiqueta, a, b in ((B_T_SEL, 'Coste de la selección', B_H_SEL,
                               B_E_SEL),
                              (B_T_FOR, 'Coste de las horas de quien enseña',
                               B_H_FOR, B_E_FOR),
                              (B_T_NUE,
                               'Coste de las horas de la persona nueva',
                               B_H_NUE, B_E_NUE)):
        motor.val(ws, 'A%d' % r, etiqueta, bold=True)
        motor.f(ws, 'B%d' % r,
                '=IFERROR(IF(OR($B${a}="",$B${b}=""),"",$B${a}*$B${b}),"")'
                .format(a=a, b=b), fmt=FMT_EUR, bold=True)
    motor.val(ws, 'A%d' % B_TOT_A, 'TOTAL COSTE DIRECTO (A)', bold=True)
    motor.f(ws, 'B%d' % B_TOT_A,
            '=IF(COUNT($B${a},$B${b},$B${c})=0,"",'
            'SUM($B${a},$B${b},$B${c}))'
            .format(a=B_T_SEL, b=B_T_FOR, c=B_T_NUE), fmt=FMT_EUR, bold=True)
    fila_total(ws, B_TOT_A, 'A', 'B')

    # A5 (auditoría 2026-09-04): bloque B reescrito para que B22 sea MARGEN
    # (comparable con el gasto real de A), no venta bruta del restaurante
    # entero escalada por la caída de UN turno. Ver el comentario de
    # `COSTE_BAJA` en `datos_ejemplo.py` para la cuenta completa.
    seccion(ws, 'A%d' % (B_B_CAB - 1),
            'B. MARGEN QUE SE DEJA DE GANAR MIENTRAS EL EQUIPO NO ESTÁ A '
            'PLENO RENDIMIENTO')
    cabecera(ws, B_B_CAB, [('A', 'Concepto'), ('B', 'Valor'), ('C', 'Nota')],
             altura=20)
    for r, etiqueta, valor, fmt, comentario in (
            (B_DIAS, 'Días hasta que el equipo vuelve a su ritmo',
             float(cb['dias_menor_rendimiento']), FMT_ENT,
             'Desde que se va la persona hasta que la nueva rinde como el '
             'puesto pide.'),
            (B_PCT, 'Peso de esa persona en la venta diaria del restaurante (%)',
             cb['pct_peso_persona_venta'], FMT_PCT,
             'NO es la caída del turno entero: es cuánto pesa ESA PERSONA en '
             'la venta de TODO el restaurante. Con una plantilla de 12 y '
             'puestos de peso desigual, 5 % es conservador para un puesto no '
             'crítico; súbelo para un puesto que sí lo sea.'),
            (B_VENTA, 'Margen medio de un día (tras prime cost)',
             cb['margen_dia_medio'], FMT_EUR,
             'Sin IVA. NO es la venta bruta: ya lleva descontado el prime '
             'cost (materia prima + personal), con el mismo % que calcula el '
             'libro 1 de este pack, «Cuadro de mando semanal»!Semana, fila '
             'TOTAL/MEDIA (37,2 % de margen sobre 3.370 €/día de venta).')):
        motor.val(ws, 'A%d' % r, etiqueta)
        motor.val(ws, 'B%d' % r, valor, fmt=fmt, verde_=True)
        motor.val(ws, 'C%d' % r, comentario, wrap=True)
        ws.row_dimensions[r].height = 30
    motor.val(ws, 'A%d' % B_TOT_B, 'MARGEN QUE NO SE GANA (B)', bold=True)
    motor.f(ws, 'B%d' % B_TOT_B,
            '=IFERROR(IF(OR($B${d}="",$B${p}="",$B${v}=""),"",'
            '$B${d}*$B${p}*$B${v}),"")'
            .format(d=B_DIAS, p=B_PCT, v=B_VENTA), fmt=FMT_EUR, bold=True)
    fila_total(ws, B_TOT_B, 'A', 'B')

    seccion(ws, 'A%d' % (B_C_CAB - 1), 'C. IMPACTO ESTIMADO DE LA BAJA')
    cabecera(ws, B_C_CAB, [('A', 'Concepto'), ('B', 'Valor'), ('C', 'Nota')],
             altura=20)
    motor.val(ws, 'A%d' % B_REF_A, 'Coste directo (A)')
    motor.f(ws, 'B%d' % B_REF_A, '=IF($B${a}="","",$B${a})'.format(a=B_TOT_A),
            fmt=FMT_EUR)
    motor.val(ws, 'A%d' % B_REF_B, 'Margen que no se gana (B)')
    motor.f(ws, 'B%d' % B_REF_B, '=IF($B${b}="","",$B${b})'.format(b=B_TOT_B),
            fmt=FMT_EUR)
    motor.val(ws, 'A%d' % B_TOTAL, 'IMPACTO ESTIMADO TOTAL (A + B)', bold=True)
    motor.f(ws, 'B%d' % B_TOTAL,
            '=IF(COUNT($B${a},$B${b})=0,"",SUM($B${a},$B${b}))'
            .format(a=B_REF_A, b=B_REF_B), fmt=FMT_EUR, bold=True)
    fila_total(ws, B_TOTAL, 'A', 'B')

    nota(ws, B_TOTAL + 2,
         'A y B SÍ son la misma clase de número: los dos son euros que '
         'salen o dejan de entrar en la cuenta de resultados. A es gasto '
         'directo (selección y formación). B ya no es venta bruta: es el '
         'MARGEN tras prime cost que esa venta habría dejado, así que '
         'sumarlos aquí sí es comparar euros con euros.',
         alto=44, wrap=True, col='A')
    ws.merge_cells('A%d:C%d' % (B_TOTAL + 2, B_TOTAL + 2))
    # M8 (auditoría 2026-09-04): «nómina» (planilla, en el uso de otros
    # países) es vocabulario de España; primera y única aparición del
    # término en este libro.
    nota(ws, B_TOTAL + 4,
         'Ninguna de estas nueve cifras es una referencia del sector: son un '
         'ejemplo. El dato bueno es el tuyo, y sale de tus nóminas (planilla, '
         'en el uso de otros países), de tu registro de jornada y de tu '
         'TPV.', alto=30, wrap=True)
    ws.merge_cells('A%d:C%d' % (B_TOTAL + 4, B_TOTAL + 4))

    motor.dv_numerica(ws, ['B%d' % r for r in (B_H_SEL, B_H_FOR, B_H_NUE,
                                               B_DIAS)],
                      minimo=0, titulo='Horas o días')
    motor.dv_numerica(ws, ['B%d' % r for r in (B_E_SEL, B_E_FOR, B_E_NUE,
                                               B_VENTA)],
                      minimo=0, titulo='Importe (€)')
    motor.dv_porcentaje(ws, ['B%d' % B_PCT], titulo='Porcentaje',
                        prompt='Se escribe en tanto por uno: 0,35 = 35 %.')
    pagina(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
def mapa():
    celdas_m = {
        'Nivel mínimo para dar una estación por cubierta': 'B%d' % M_CUBRE,
        'Nivel a partir del cual se puede enseñar': 'B%d' % M_ENSENA,
        'Umbral de punto único de fallo': 'B%d' % M_RIESGO,
    }
    for j, est in enumerate(D.ESTACIONES):
        celdas_m['Cabecera de la estación «%s»' % est] = \
            '%s%d' % (get_column_letter(COL_EST0 + j), M_CAB)
    for i, p in enumerate(D.PLANTILLA):
        r = M_INI + i
        celdas_m['Estaciones que puede sostener %s (%s)' % (p[1], p[0])] = \
            '%s%d' % (COL_SOST, r)
        celdas_m['Estaciones que puede enseñar %s (%s)' % (p[1], p[0])] = \
            '%s%d' % (COL_ENS, r)

    celdas_c = {
        'Nivel mínimo para dar por cubierta (espejo)': 'C%d' % C_CUBRE,
        'Nivel para enseñar (espejo)': 'C%d' % C_ENSENA,
        'Umbral de punto único de fallo (espejo)': 'C%d' % C_RIESGO,
        'Estaciones registradas': 'C%d' % (C_RES + 1),
        'Personas registradas en la matriz': 'C%d' % (C_RES + 2),
        'Estaciones en riesgo': 'C%d' % (C_RES + 3),
        'Personas de la estación mejor cubierta': 'C%d' % (C_RES + 4),
        'Personas de la estación peor cubierta': 'C%d' % (C_RES + 5),
        'Estaciones en riesgo (%)': 'C%d' % (C_RES + 6),
    }
    for j, est in enumerate(D.ESTACIONES):
        r = C_INI + j
        celdas_c['Nombre de la estación %d' % (j + 1)] = 'A%d' % r
        celdas_c['Personas que sostienen «%s»' % est] = 'B%d' % r
        celdas_c['Personas que pueden enseñar «%s»' % est] = 'C%d' % r
        celdas_c['Cobertura de «%s» (%%)' % est] = 'E%d' % r
        celdas_c['Alerta de «%s»' % est] = 'F%d' % r

    celdas_x = {
        'Fecha de corte del plan': 'C%d' % X_CORTE,
        'Días de aviso': 'C%d' % X_AVISO,
        'Acciones registradas': 'C%d' % (X_RES + 1),
        'Acciones completadas': 'C%d' % (X_RES + 2),
        'Acciones en curso': 'C%d' % (X_RES + 3),
        'Acciones planificadas': 'C%d' % (X_RES + 4),
        'Acciones vencidas': 'C%d' % (X_RES + 5),
        'Acciones que vencen pronto': 'C%d' % (X_RES + 6),
        'Acciones completadas (%)': 'C%d' % (X_RES + 7),
    }
    for i, accion in enumerate(D.PLAN_CROSS_TRAINING):
        r = X_INI + i
        etiqueta = '%s hacia «%s»' % (accion[0], accion[1])
        celdas_x['Niveles que faltan: ' + etiqueta] = 'J%d' % r
        celdas_x['Días hasta la fecha: ' + etiqueta] = 'K%d' % r
        celdas_x['Situación del plazo: ' + etiqueta] = 'L%d' % r

    celdas_b = {
        'Horas de selección': 'B%d' % B_H_SEL,
        'Coste por hora de quien selecciona': 'B%d' % B_E_SEL,
        'Coste de la selección': 'B%d' % B_T_SEL,
        'Horas de quien enseña': 'B%d' % B_H_FOR,
        'Coste por hora de quien enseña': 'B%d' % B_E_FOR,
        'Coste de las horas de quien enseña': 'B%d' % B_T_FOR,
        'Horas de formación de la persona nueva': 'B%d' % B_H_NUE,
        'Coste por hora de la persona nueva': 'B%d' % B_E_NUE,
        'Coste de las horas de la persona nueva': 'B%d' % B_T_NUE,
        'Total del coste directo (A)': 'B%d' % B_TOT_A,
        'Días de menor rendimiento': 'B%d' % B_DIAS,
        'Peso de esa persona en la venta diaria del restaurante':
            'B%d' % B_PCT,
        'Margen medio de un día (tras prime cost)': 'B%d' % B_VENTA,
        'Margen que no se gana (B)': 'B%d' % B_TOT_B,
        'Impacto estimado total de la baja': 'B%d' % B_TOTAL,
    }
    return {
        'fichero': NOMBRE + '.xlsx',
        'producto': 'manual-manager-restaurante',
        'punto_unico_de_fallo': 'Fríos y entrantes',
        'hojas': {
            'Matriz': {
                'celdas': celdas_m,
                'tablas': [
                    {'titulo': 'Leyenda de niveles',
                     'cols': [['Nivel', 'A', 'num'],
                              ['Qué significa', 'B', 'txt']],
                     'filas': [M_LEY_INI, M_LEY_FIN]},
                    {'titulo': 'Matriz de polivalencia (12 personas x 6 '
                               'estaciones sembradas)',
                     'cols': ([['Id', 'A', 'txt'], ['Nombre', 'B', 'txt'],
                               ['Puesto', 'C', 'txt'],
                               ['Estación principal', 'D', 'txt']]
                              + [[D.ESTACIONES[j],
                                  get_column_letter(COL_EST0 + j), 'num']
                                 for j in range(len(D.ESTACIONES))]
                              + [['Estaciones que puede sostener', COL_SOST,
                                  'num'],
                                 ['Estaciones que puede enseñar', COL_ENS,
                                  'num']]),
                     'filas': [M_INI, M_INI + len(D.PLANTILLA) - 1]},
                ],
            },
            'Plan de Cross-Training': {
                'celdas': celdas_x,
                'tablas': [
                    {'titulo': 'Plan de cross-training',
                     'cols': [['#', 'A', 'num'],
                              ['Id del empleado', 'B', 'txt'],
                              ['Nombre', 'C', 'txt'],
                              ['Estación objetivo', 'D', 'txt'],
                              ['Nivel actual', 'E', 'num'],
                              ['Nivel objetivo', 'F', 'num'],
                              ['Quién le enseña', 'G', 'txt'],
                              ['Fecha objetivo', 'H', 'txt'],
                              ['Estado', 'I', 'txt'],
                              ['Niveles que faltan', 'J', 'num'],
                              ['Días hasta la fecha', 'K', 'num'],
                              ['Situación del plazo', 'L', 'txt']],
                     'filas': [X_INI, X_INI + len(D.PLAN_CROSS_TRAINING) - 1]},
                ],
            },
            'Cobertura por Estación': {
                'celdas': celdas_c,
                'tablas': [
                    {'titulo': 'Cobertura por estación',
                     'cols': [['Estación', 'A', 'txt'],
                              ['Personas que la sostienen', 'B', 'num'],
                              ['Personas que pueden enseñarla', 'C', 'num'],
                              ['Personas en la matriz', 'D', 'num'],
                              ['Cobertura (%)', 'E', 'pct1'],
                              ['Alerta', 'F', 'txt']],
                     'filas': [C_INI, C_INI + len(D.ESTACIONES) - 1]},
                ],
            },
            'Coste de una Baja': {
                'celdas': celdas_b,
                'tablas': [
                    {'titulo': 'Coste directo de cubrir la baja',
                     'cols': [['Concepto', 'A', 'txt'], ['Valor', 'B', 'eur'],
                              ['Nota', 'C', 'txt']],
                     'filas': [B_H_SEL, B_TOT_A]},
                    {'titulo': 'Venta que se deja de hacer',
                     'cols': [['Concepto', 'A', 'txt'], ['Valor', 'B', 'eur'],
                              ['Nota', 'C', 'txt']],
                     'filas': [B_DIAS, B_TOT_B]},
                ],
            },
        },
    }


def main():
    wb = Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_matriz(wb)
    hoja_plan(wb)
    hoja_cobertura(wb)
    hoja_coste_baja(wb)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO
    wb.properties.subject = PRODUCTO + ' · v1.0'

    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        motor.proteger(ws)

    destino = os.path.join(AQUI, 'build')
    os.makedirs(destino, exist_ok=True)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')
    wb.save(ruta)
    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)
    print('escrito:', ruta)
    print('formulas registradas:', len(motor.REGISTRO))


if __name__ == '__main__':
    main()
