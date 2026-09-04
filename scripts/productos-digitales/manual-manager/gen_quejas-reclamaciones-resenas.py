#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_quejas-reclamaciones-resenas.py — Libro 3 del pack «Manual del Manager de
Restaurante» (SPEC §2.2, fila 3).

Genera `build/quejas-reclamaciones-resenas.xlsx`:

  Instrucciones · Parámetros · Registro de Quejas · Reclamaciones Formales ·
  Reseñas · Resumen

Son TRES cosas distintas y por eso son tres hojas: la queja se resuelve en
sala contra el SLA que se pone la casa; la hoja oficial de reclamaciones tiene
un plazo legal que fija cada COMUNIDAD; la reseña es pública y se responde.
El libro las registra por separado y las cruza en la hoja «Resumen», que es
donde se ve si las quejas se repiten por el MISMO motivo (entonces es un
problema de proceso, no de una persona) y si la reclamación formal se contestó
dentro del plazo de tu comunidad.

DECISIONES TÉCNICAS
-------------------
* Cero constantes dentro de una fórmula: el SLA por gravedad, el plazo legal
  por comunidad, la conversión de días a horas y los objetivos de la casa
  viven en celdas. El SLA y el plazo se recuperan con INDEX/MATCH.
* Funciones prohibidas (INDIRECT, COUNTA, PMT, OFFSET, XLOOKUP, LET, LAMBDA,
  matrices dinámicas y RANK — que pycel no implementa): cero.
* «Sin dato» = `""`, nunca `0`; `IFERROR(...,"")` en todo cociente.
* Los semáforos comparan con `ISNUMBER` y SIEMPRE contra celdas de la MISMA
  hoja: el formato condicional de Excel no puede referirse a otra hoja.
* Los plazos autonómicos se siembran SOLO con lo verificado contra la norma
  (Cataluña y Andalucía). El resto queda vacío y editable: inventarlos sería
  peor que dejarlos en blanco.

Salida fija (sin argumentos): `<carpeta>/build/quejas-reclamaciones-resenas.xlsx`
"""
import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(
    0, '/Users/johnguerrero/chefpro-modernize/scripts/productos-digitales/'
       'guias-v2_0')
import motor  # noqa: E402

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import datos_ejemplo as DE  # noqa: E402

PRODUCTO = 'manual-manager-restaurante'
motor.CTX['producto'] = PRODUCTO

# --------------------------------------------------------------------------
NOMBRE = 'quejas-reclamaciones-resenas'
TITULO_LIBRO = 'Quejas, Reclamaciones Formales y Reseñas'
SUBTITULO = 'AI Chef Pro · aichef.pro — Manual del Manager de Restaurante'
SUBJECT = 'Manual del Manager de Restaurante · v1.0 · septiembre 2026'
VERSION = DE.VERSION_LINE
BIO = DE.BIO
DESPROTEGER = ('Para editar la estructura o una celda que no esté en verde: '
               'Revisar, Desproteger hoja (no tiene contraseña).')
LEYENDA_VERDE = 'Celdas verdes = campos editables'
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'

GOLD, GRIS = 'FFD700', '888888'
CAB_BG, CAB_FG = '2D2D2D', 'FFFFFF'
CREMA, AZUL, GRIS_BG = 'FFF8E1', '1565C0', 'F2F2F2'
EUR, PCT, ENT = motor.FMT_EUR, motor.FMT_PCT, motor.FMT_ENT
FECHA = motor.FMT_FECHA
DEC1 = '#,##0.0'
DEC2 = '#,##0.00'
MES = 'mmm yyyy'

VERIF = 'Verificado el 04-09-2026'

# --- filas de la hoja «Parámetros» ----------------------------------------
P_SLA_CAB = 5
P_SLA_INI = 6
P_SLA_FIN = 8
P_HORAS = 11                     # horas que tiene un día (conversión)
P_PLZ_CAB = 14
P_PLZ_INI = 15
P_PLZ_FIN = 17
P_MOT_CAB = 20
P_MOT_INI = 21
P_MOT_FIN = P_MOT_INI + len(DE.MOTIVOS_QUEJA) - 1
P_CAN_CAB = P_MOT_FIN + 3
P_CAN_INI = P_CAN_CAB + 1
P_CAN_FIN = P_CAN_INI + len(DE.CANALES_QUEJA) - 1

# --- filas de las hojas de registro ---------------------------------------
Q_CAB, Q_INI = 4, 5
Q_FIN = 54                       # 30 sembradas + 20 libres
R_CAB, R_INI = 4, 5
R_FIN = 24                       # 3 sembradas + 17 libres
S_CAB, S_INI = 4, 5
S_FIN = 64                       # 40 sembradas + 20 libres

# --- filas de la hoja «Resumen» -------------------------------------------
X_OBJ_SLA = 5
X_OBJ_RESP = 6
X_OBJ_ESTRELLAS = 7
X_TOLERANCIA = 8
X_Q_INI = 11                     # quejas registradas
X_MOT_CAB = 21
X_MOT_INI = 22
X_MOT_FIN = X_MOT_INI + len(DE.MOTIVOS_QUEJA) - 1
X_GRA_CAB = X_MOT_FIN + 3
X_GRA_INI = X_GRA_CAB + 1
X_GRA_FIN = X_GRA_INI + 2
X_REC_INI = X_GRA_FIN + 3
X_RES_INI = X_REC_INI + 6
X_MES_CAB = X_RES_INI + 6
X_MES_INI = X_MES_CAB + 1
X_MES_FIN = X_MES_INI + 11       # 6 meses sembrados + 6 libres

PLATAFORMAS = ['Google', 'TripAdvisor', 'TheFork', 'Facebook', 'Otra']
TEMAS = ['Comida', 'Servicio', 'Ambiente', 'Espera', 'Precio', 'Reserva',
         'Otro']
SI_NO = ['Sí', 'No']
RESPONSABLES = ['%s · %s' % (p[0], p[1]) for p in DE.PLANTILLA]
COMUNIDADES = [c[0] for c in DE.PARAMETROS_QUEJAS['plazos_autonomicos']]

# Plazo en días y tipo de día de cada comunidad VERIFICADA. Sólo dos: el resto
# queda vacío y editable (SPEC §2.2, fila 3).
PLAZOS = [
    ('Cataluña', '1 mes', 30, 'Naturales',
     VERIF + ' · Decret 121/2013 y Codi de consum de Cataluña (Llei 22/2010), '
     'art. 126-9: hay que contestar en el plazo máximo de 1 mes desde la '
     'presentación de la hoja · '
     'https://www.boe.es/buscar/act.php?id=BOE-A-2010-13115'),
    ('Andalucía', '10 días hábiles', 10, 'Hábiles',
     VERIF + ' · Decreto 82/2022, arts. 4, 7 y 12: respuesta por escrito en '
     '10 días hábiles, cartel de tamaño mínimo DIN-A4 con letra de 0,7 cm o '
     'más, y hoja electrónica obligatoria desde mayo de 2026 · '
     'https://www.juntadeandalucia.es/boja/2022/95/42'),
    ('Otras comunidades', '', None, '',
     'Consulta tu comunidad: la hoja oficial de reclamaciones es competencia '
     'AUTONÓMICA y el plazo cambia de una a otra. Escribe aquí el tuyo, con '
     'la norma que lo fija, y el libro lo usará en la hoja «Reclamaciones '
     'Formales».'),
]

MOTIVO_NOTA = {
    'Espera excesiva': 'Tiempo de espera para sentarse, para la comanda o '
                       'para que salga el plato.',
    'Plato frío o mal temperado': 'Temperatura de servicio fuera del estándar '
                                  'de la ficha.',
    'Error en la comanda': 'Plato distinto del pedido, modificación perdida o '
                           'guarnición cambiada.',
    'Trato del personal': 'Forma, tono o atención de una persona del equipo.',
    'Cobro incorrecto': 'Precio, cantidad o forma de pago mal aplicados.',
    'Limpieza': 'Sala, aseos, mantelería, cristalería o cubertería.',
    'Reserva no encontrada': 'La reserva no consta, o consta con otros datos.',
    'Ruido': 'Volumen de música, acústica de la sala o ruido de office.',
    'Información de alérgenos': 'Información incompleta, tardía o incorrecta '
                                'sobre los 14 alérgenos.',
    'Producto en mal estado': 'Producto en mal estado o sospecha de incidente '
                              'alimentario. Siempre gravedad 3.',
}


# --------------------------------------------------------------------------
def cabecera(ws, titulo):
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16, color=GOLD)
    motor.val(ws, 'A2', SUBTITULO)
    ws['A2'].font = Font(size=10, color=GRIS)


def apunte(ws, coord, texto):
    motor.val(ws, coord, texto, wrap=True)
    ws[coord].font = Font(size=9, color=GRIS)


def setup(ws, apaisado=True, titulos=None):
    ws.page_setup.paperSize = 9                      # A4
    ws.page_setup.orientation = 'landscape' if apaisado else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.39
    ws.page_margins.top = ws.page_margins.bottom = 0.59
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if titulos:
        ws.print_title_rows = titulos


def encabezados(ws, fila, cols, alto=38):
    for letra, texto, ancho in cols:
        cel = ws[letra + str(fila)]
        cel.value = texto
        cel.font = Font(bold=True, color=CAB_FG)
        cel.fill = PatternFill('solid', fgColor=CAB_BG)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho
    ws.row_dimensions[fila].height = alto


def seccion(ws, coord, texto):
    motor.val(ws, coord, texto, bold=True)
    ws[coord].font = Font(bold=True, size=12, color=GOLD)


def etiqueta_valor(ws, fila, etiqueta, coord_valor, fmt=None, formula=None,
                   valor=None, verde_=False, destacar=False, nota=None,
                   col_nota='F'):
    motor.val(ws, 'A' + str(fila), etiqueta)
    if formula is not None:
        cel = motor.f(ws, coord_valor, formula, fmt=fmt, bold=destacar)
    else:
        cel = motor.val(ws, coord_valor, valor, fmt=fmt, verde_=verde_,
                        bold=destacar)
    if destacar:
        cel.fill = PatternFill('solid', fgColor=CREMA)
    if nota:
        motor.val(ws, col_nota + str(fila), nota, wrap=True)
    return cel


# --------------------------------------------------------------------------
# Hoja «Instrucciones»
# --------------------------------------------------------------------------
PASOS = [
    '1. Abre «Parámetros» y pon TU SLA: cuántas horas te das para cerrar una '
    'queja leve, una media y una grave. Es un compromiso de la casa, no una '
    'norma: por eso está en celda verde.',
    '2. En la misma hoja, busca tu comunidad en la tabla de plazos legales. '
    'Sólo vienen sembradas Cataluña y Andalucía, que son las dos verificadas '
    'contra la norma. Si la tuya no está, escribe su plazo y su norma en la '
    'fila «Otras comunidades»: el libro la usará igual.',
    '3. Registra cada queja en «Registro de Quejas» según ocurre. Fecha, '
    'canal, motivo de la lista, gravedad de 1 a 3, quién la atiende, qué se '
    'hizo y la fecha en que quedó cerrada. El libro calcula solo las horas '
    'que tardaste, el SLA que le tocaba y si lo cumpliste.',
    '4. Una hoja OFICIAL de reclamaciones no es una queja: va a '
    '«Reclamaciones Formales», con su número, la comunidad y la fecha en que '
    'contestaste por escrito. El libro trae el plazo legal de esa comunidad y '
    'te dice si llegaste a tiempo.',
    '5. Vuelca las reseñas en «Reseñas»: plataforma, fecha, estrellas, de qué '
    'va y si está respondida. El mes lo calcula el libro.',
    '6. Lee «Resumen» una vez al mes. Ahí está lo que importa: cuál es el '
    'motivo que más se repite, cuánto tardas de media en cerrar, qué '
    'porcentaje de SLA cumples, si alguna reclamación se contestó fuera de '
    'plazo y cómo se mueve la media de estrellas mes a mes.',
    '7. Si un motivo concentra las quejas, deja de tratarlo como incidentes '
    'sueltos: es un problema de proceso. Llévalo al plan de acción del libro '
    'de reuniones y ponle responsable y fecha.',
]

NOTAS = [
    'QUEJA, RECLAMACIÓN Y RESEÑA NO SON LO MISMO, y por eso son tres hojas. '
    'La queja se resuelve en sala y se mide contra el SLA que se pone la casa. '
    'La hoja oficial de reclamaciones es un procedimiento administrativo con '
    'un plazo de respuesta que fija tu comunidad autónoma. La reseña es '
    'pública, la lee quien todavía no ha venido, y se responde.',
    'La hoja oficial de reclamaciones es competencia AUTONÓMICA: el modelo, '
    'el cartel obligatorio y el plazo de respuesta cambian de una comunidad a '
    'otra. Este libro sólo siembra las dos comunidades cuyo plazo se ha '
    'verificado contra la norma, con su fecha de verificación y su URL. '
    'Para el resto, la celda está vacía a propósito.',
    'Los días que cuenta el libro son NATURALES. Cuando tu comunidad cuenta '
    'días HÁBILES, la columna «Se cuentan» lo dice y la comparación queda del '
    'lado seguro: te avisará antes de tiempo, nunca después. Para un '
    'expediente real, cuenta los hábiles de tu calendario.',
    'Las 30 quejas, las 3 reclamaciones y las 40 reseñas sembradas son un '
    'EJEMPLO del restaurante modelado del pack. Bórralas y pon las tuyas: el '
    'libro sigue funcionando con las filas vacías.',
    'El SLA es tuyo, no de la ley. Ponlo donde puedas cumplirlo: un SLA que '
    'incumples el 40 por ciento de las veces no mide nada.',
    'Una gravedad 3 (riesgo sanitario, legal o reputacional) no se cierra sin '
    'respuesta escrita. Si además pide la hoja oficial, esa misma incidencia '
    'tiene que aparecer en las DOS hojas: en el registro de quejas y en el de '
    'reclamaciones formales.',
]


def hoja_instrucciones(wb):
    ws = wb.create_sheet('Instrucciones', 0)
    ws.column_dimensions['A'].width = 118.0
    cabecera(ws, TITULO_LIBRO)
    motor.val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12, color=GOLD)
    fila = 5
    for paso in PASOS:
        motor.val(ws, 'A' + str(fila), paso, wrap=True)
        fila += 1
    fila += 1
    motor.val(ws, 'A' + str(fila), LEYENDA_VERDE, verde_=True)
    fila += 2
    for nota in NOTAS:
        motor.val(ws, 'A' + str(fila), nota, wrap=True)
        fila += 2
    motor.val(ws, 'A' + str(fila),
              'Este libro NO sustituye al asesoramiento jurídico ni al '
              'expediente de consumo. Es el registro operativo que te permite '
              'llegar a ese expediente con las fechas puestas.', wrap=True)
    fila += 2
    motor.val(ws, 'A' + str(fila), DESPROTEGER, wrap=True)
    motor.val(ws, 'A' + str(fila + 1), BIO, wrap=True)
    motor.val(ws, 'A' + str(fila + 2), VERSION, wrap=True)
    setup(ws, apaisado=False)
    return ws


# --------------------------------------------------------------------------
# Hoja «Parámetros»
# --------------------------------------------------------------------------
def hoja_parametros(wb):
    ws = wb.create_sheet('Parámetros')
    cabecera(ws, 'Parámetros — el SLA es tuyo; el plazo legal, de tu comunidad')

    seccion(ws, 'A4', 'SLA DE LA CASA POR GRAVEDAD DE LA QUEJA')
    encabezados(ws, P_SLA_CAB, [
        ('A', 'Gravedad (1-3)', 13), ('B', 'Qué significa', 56),
        ('C', 'SLA de cierre (horas)', 14), ('D', '', 4),
        ('E', '', 4), ('F', 'Notas y fuente', 90),
    ])
    for i, (g, texto) in enumerate(DE.PARAMETROS_QUEJAS['escala_gravedad']):
        r = P_SLA_INI + i
        motor.val(ws, 'A%d' % r, g, fmt=ENT)
        motor.val(ws, 'B%d' % r, texto, wrap=True)
        motor.val(ws, 'C%d' % r,
                  DE.PARAMETROS_QUEJAS['sla_horas_por_gravedad'][g], fmt=ENT,
                  verde_=True)
    motor.dv_numerica(ws, ['C%d' % r for r in range(P_SLA_INI, P_SLA_FIN + 1)],
                      minimo=1, maximo=720, titulo='SLA en horas',
                      mensaje='Escribe las horas de las que te quieres dar '
                              'para cerrar una queja de esa gravedad '
                              '(entre 1 y 720).')
    motor.val(ws, 'F%d' % P_SLA_INI,
              'El SLA es un compromiso de la casa, no una obligación legal. '
              'La escala de gravedad es la del capítulo 17 del manual.',
              wrap=True)

    seccion(ws, 'A10', 'CONSTANTE DE CÁLCULO')
    motor.val(ws, 'A%d' % P_HORAS, 'Horas que tiene un día')
    cel = motor.val(ws, 'C%d' % P_HORAS, 24, fmt=ENT)
    cel.fill = PatternFill('solid', fgColor=GRIS_BG)
    motor.val(ws, 'F%d' % P_HORAS,
              'Excel resta fechas en DÍAS. Esta celda convierte esa resta en '
              'horas para poder compararla con el SLA. No es un parámetro de '
              'tu negocio: por eso no está en verde.', wrap=True)

    seccion(ws, 'A13', 'PLAZO LEGAL DE RESPUESTA A LA HOJA OFICIAL DE '
                       'RECLAMACIONES, POR COMUNIDAD')
    encabezados(ws, P_PLZ_CAB, [
        ('A', 'Comunidad', 22), ('B', 'Plazo legal', 18),
        ('C', 'Días para el aviso', 14), ('D', 'Se cuentan', 12),
        ('E', '', 4), ('F', 'Norma, verificación y URL', 90),
    ])
    for i, (com, plazo, dias, tipo, nota) in enumerate(PLAZOS):
        r = P_PLZ_INI + i
        motor.val(ws, 'A%d' % r, com)
        motor.val(ws, 'B%d' % r, plazo, verde_=True)
        motor.val(ws, 'C%d' % r, dias, fmt=ENT, verde_=True)
        motor.val(ws, 'D%d' % r, tipo, verde_=True)
        motor.val(ws, 'F%d' % r, nota, wrap=True)
    motor.dv_numerica(ws, ['C%d' % r for r in range(P_PLZ_INI, P_PLZ_FIN + 1)],
                      minimo=1, maximo=365, titulo='Plazo en días',
                      mensaje='Escribe el plazo de respuesta de tu comunidad '
                              'en días (entre 1 y 365).')
    motor.dv_lista(ws, ['D%d' % r for r in range(P_PLZ_INI, P_PLZ_FIN + 1)],
                   ['Naturales', 'Hábiles'], titulo='Se cuentan',
                   mensaje='Indica si tu comunidad cuenta días naturales o '
                           'días hábiles.')

    seccion(ws, 'A%d' % (P_MOT_CAB - 1),
            'MOTIVOS DE QUEJA (lista cerrada del registro)')
    encabezados(ws, P_MOT_CAB, [
        ('A', 'Motivo', None), ('B', 'Qué recoge', None),
    ], alto=26)
    for i, m in enumerate(DE.MOTIVOS_QUEJA):
        r = P_MOT_INI + i
        motor.val(ws, 'A%d' % r, m)
        motor.val(ws, 'B%d' % r, MOTIVO_NOTA.get(m, ''), wrap=True)
    motor.val(ws, 'F%d' % P_MOT_INI,
              'La lista es cerrada A PROPÓSITO: si cada queja se escribe con '
              'palabras distintas, el resumen no puede contar cuál se repite, '
              'que es justo lo que hay que ver.', wrap=True)

    seccion(ws, 'A%d' % (P_CAN_CAB - 1), 'CANALES DE ENTRADA')
    encabezados(ws, P_CAN_CAB, [('A', 'Canal', None)], alto=26)
    for i, c in enumerate(DE.CANALES_QUEJA):
        motor.val(ws, 'A%d' % (P_CAN_INI + i), c)

    motor.val(ws, 'A%d' % (P_CAN_FIN + 2),
              'La hoja oficial de reclamaciones es competencia AUTONÓMICA. '
              'Este libro sólo siembra los dos plazos verificados contra la '
              'norma el 04-09-2026. Antes de fiarte de cualquier otro dato, '
              'compruébalo en el boletín de tu comunidad.', wrap=True)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
# Hoja «Registro de Quejas»
# --------------------------------------------------------------------------
def hoja_quejas(wb):
    ws = wb.create_sheet('Registro de Quejas')
    cabecera(ws, 'Registro de quejas — lo que se resuelve en sala')
    apunte(ws, 'G3', 'Las horas hasta el cierre, el SLA de esa gravedad y el '
                     'cumplimiento los calcula el libro.')
    encabezados(ws, Q_CAB, [
        ('A', '#', 5), ('B', 'Fecha de la queja', 13),
        ('C', 'Canal', 18), ('D', 'Motivo', 24),
        ('E', 'Gravedad (1-3)', 10), ('F', 'Responsable', 18),
        ('G', 'Acción tomada', 54), ('H', 'Fecha de cierre', 13),
        ('I', 'Horas hasta el cierre', 12),
        ('J', 'SLA de esa gravedad (horas)', 12),
        ('K', 'SLA cumplido', 12), ('L', 'Notas', 34),
    ])
    ws.freeze_panes = 'C5'

    ids = {p[0]: '%s · %s' % (p[0], p[1]) for p in DE.PLANTILLA}
    v_canal, v_motivo, v_resp, v_grav, v_fecha = [], [], [], [], []
    for i in range(Q_FIN - Q_INI + 1):
        r = Q_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.QUEJAS):
            fecha, canal, motivo, grav, resp, accion, cierre = DE.QUEJAS[i]
            motor.val(ws, 'B%d' % r, DE._fecha(fecha), fmt=FECHA)
            motor.val(ws, 'C%d' % r, canal)
            motor.val(ws, 'D%d' % r, motivo)
            motor.val(ws, 'E%d' % r, grav, fmt=ENT)
            motor.val(ws, 'F%d' % r, ids.get(resp, resp))
            motor.val(ws, 'G%d' % r, accion, wrap=True)
            motor.val(ws, 'H%d' % r, DE._fecha(cierre), fmt=FECHA)
        else:
            ws['B%d' % r].number_format = FECHA
            ws['E%d' % r].number_format = ENT
            ws['H%d' % r].number_format = FECHA
        motor.verde(ws, 'B%d:H%d' % (r, r))
        motor.verde(ws, 'L%d' % r)
        v_fecha += ['B%d' % r, 'H%d' % r]
        v_canal.append('C%d' % r)
        v_motivo.append('D%d' % r)
        v_grav.append('E%d' % r)
        v_resp.append('F%d' % r)
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF(OR($B{r}="",$H{r}=""),"",'
                '($H{r}-$B{r})*Parámetros!$C${h}),"")'.format(r=r, h=P_HORAS),
                fmt=ENT)
        motor.f(ws, 'J%d' % r,
                '=IFERROR(IF($E{r}="","",INDEX(Parámetros!$C${a}:$C${b},'
                'MATCH($E{r},Parámetros!$A${a}:$A${b},0))),"")'
                .format(r=r, a=P_SLA_INI, b=P_SLA_FIN), fmt=ENT)
        motor.f(ws, 'K%d' % r,
                '=IFERROR(IF(OR($I{r}="",$J{r}=""),"",'
                'IF($I{r}<=$J{r},"Sí","No")),"")'.format(r=r), bold=True)

    motor.dv_lista(ws, v_canal, DE.CANALES_QUEJA, titulo='Canal no válido')
    motor.dv_lista(ws, v_motivo, DE.MOTIVOS_QUEJA, titulo='Motivo no válido')
    motor.dv_lista(ws, v_resp, RESPONSABLES, titulo='Responsable no válido',
                   mensaje='Elige a alguien de la plantilla. Si tu equipo es '
                           'otro, desprotege la hoja y cambia la lista.')
    motor.dv_numerica(ws, v_grav, minimo=1, maximo=3, titulo='Gravedad',
                      mensaje='1 leve, 2 media, 3 grave. La escala está en la '
                              'hoja «Parámetros».')
    motor.dv_fecha(ws, v_fecha)
    motor.semaforo_texto(ws, 'K%d:K%d' % (Q_INI, Q_FIN), (
        ('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    motor.val(ws, 'A%d' % (Q_FIN + 2),
              'Una queja sin fecha de cierre no cuenta como incumplida: '
              'cuenta como ABIERTA, y por eso su fila se queda vacía. Cierra '
              'la fecha cuando la incidencia esté resuelta de verdad.',
              wrap=True)
    setup(ws, titulos='$%d:$%d' % (Q_CAB, Q_CAB))
    return ws


# --------------------------------------------------------------------------
# Hoja «Reclamaciones Formales»
# --------------------------------------------------------------------------
def hoja_reclamaciones(wb):
    ws = wb.create_sheet('Reclamaciones Formales')
    cabecera(ws, 'Reclamaciones formales — la hoja oficial y su plazo legal')
    apunte(ws, 'F3', 'El plazo sale de la tabla de tu comunidad en '
                     '«Parámetros». Los días que cuenta el libro son '
                     'NATURALES.')
    encabezados(ws, R_CAB, [
        ('A', '#', 5), ('B', 'Fecha de entrega de la hoja', 14),
        ('C', 'Número de hoja', 18), ('D', 'Comunidad', 20),
        ('E', 'Fecha de respuesta por escrito', 14),
        ('F', 'Días naturales transcurridos', 12),
        ('G', 'Plazo de la comunidad (días)', 12),
        ('H', 'Se cuentan', 12), ('I', 'Dentro de plazo', 13),
        ('J', 'Objeto de la reclamación y respuesta', 60),
    ])
    ws.freeze_panes = 'C5'

    v_com, v_fecha = [], []
    for i in range(R_FIN - R_INI + 1):
        r = R_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.RECLAMACIONES):
            entrega, num, com, respuesta = DE.RECLAMACIONES[i]
            motor.val(ws, 'B%d' % r, DE._fecha(entrega), fmt=FECHA)
            motor.val(ws, 'C%d' % r, num)
            motor.val(ws, 'D%d' % r, com)
            motor.val(ws, 'E%d' % r, DE._fecha(respuesta), fmt=FECHA)
        else:
            ws['B%d' % r].number_format = FECHA
            ws['E%d' % r].number_format = FECHA
        motor.verde(ws, 'B%d:E%d' % (r, r))
        motor.verde(ws, 'J%d' % r)
        v_com.append('D%d' % r)
        v_fecha += ['B%d' % r, 'E%d' % r]
        motor.f(ws, 'F%d' % r,
                '=IFERROR(IF(OR($B{r}="",$E{r}=""),"",$E{r}-$B{r}),"")'
                .format(r=r), fmt=ENT)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF($D{r}="","",IF(INDEX(Parámetros!$C${a}:$C${b},'
                'MATCH($D{r},Parámetros!$A${a}:$A${b},0))=0,"",'
                'INDEX(Parámetros!$C${a}:$C${b},'
                'MATCH($D{r},Parámetros!$A${a}:$A${b},0)))),"")'
                .format(r=r, a=P_PLZ_INI, b=P_PLZ_FIN), fmt=ENT)
        motor.f(ws, 'H%d' % r,
                '=IFERROR(IF($D{r}="","",IF(INDEX(Parámetros!$D${a}:$D${b},'
                'MATCH($D{r},Parámetros!$A${a}:$A${b},0))=0,"",'
                'INDEX(Parámetros!$D${a}:$D${b},'
                'MATCH($D{r},Parámetros!$A${a}:$A${b},0)))),"")'
                .format(r=r, a=P_PLZ_INI, b=P_PLZ_FIN))
        motor.f(ws, 'I%d' % r,
                '=IFERROR(IF(OR($F{r}="",$G{r}=""),"",'
                'IF($F{r}<=$G{r},"Sí","No")),"")'.format(r=r), bold=True)

    motor.dv_lista(ws, v_com, COMUNIDADES, titulo='Comunidad no válida',
                   mensaje='Elige tu comunidad. Si no está sembrada, usa '
                           '«Otras comunidades» y escribe su plazo en la hoja '
                           '«Parámetros».')
    motor.dv_fecha(ws, v_fecha)
    motor.semaforo_texto(ws, 'I%d:I%d' % (R_INI, R_FIN), (
        ('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    motor.val(ws, 'A%d' % (R_FIN + 2),
              'Entregar la hoja es obligatorio y gratuito, y negarse es una '
              'infracción de consumo. Lo que decide si llegas a tiempo es la '
              'RESPUESTA POR ESCRITO, no la conversación en sala.', wrap=True)
    motor.val(ws, 'A%d' % (R_FIN + 3),
              'Cuando la columna «Se cuentan» dice «Hábiles», el libro sigue '
              'comparando días naturales: te avisa antes, nunca después. Para '
              'el expediente, cuenta los hábiles de tu calendario.', wrap=True)
    setup(ws, titulos='$%d:$%d' % (R_CAB, R_CAB))
    return ws


# --------------------------------------------------------------------------
# Hoja «Reseñas»
# --------------------------------------------------------------------------
def hoja_resenas(wb):
    ws = wb.create_sheet('Reseñas')
    cabecera(ws, 'Reseñas — lo público, lo que lee quien todavía no ha venido')
    apunte(ws, 'F3', 'El mes lo calcula el libro desde la fecha: la hoja '
                     '«Resumen» agrupa por él.')
    encabezados(ws, S_CAB, [
        ('A', '#', 5), ('B', 'Plataforma', 16), ('C', 'Fecha', 13),
        ('D', 'Estrellas (1-5)', 11), ('E', 'Tema', 16),
        ('F', 'Respondida', 11), ('G', 'Mes', 13), ('H', 'Notas', 50),
    ])
    ws.freeze_panes = 'C5'

    v_plat, v_tema, v_resp, v_est, v_fecha = [], [], [], [], []
    for i in range(S_FIN - S_INI + 1):
        r = S_INI + i
        motor.val(ws, 'A%d' % r, i + 1, fmt=ENT)
        if i < len(DE.RESENAS):
            plat, fecha, estrellas, tema, respondida = DE.RESENAS[i]
            motor.val(ws, 'B%d' % r, plat)
            motor.val(ws, 'C%d' % r, DE._fecha(fecha), fmt=FECHA)
            motor.val(ws, 'D%d' % r, estrellas, fmt=ENT)
            motor.val(ws, 'E%d' % r, tema)
            motor.val(ws, 'F%d' % r, respondida)
        else:
            ws['C%d' % r].number_format = FECHA
            ws['D%d' % r].number_format = ENT
        motor.verde(ws, 'B%d:F%d' % (r, r))
        motor.verde(ws, 'H%d' % r)
        v_plat.append('B%d' % r)
        v_tema.append('E%d' % r)
        v_resp.append('F%d' % r)
        v_est.append('D%d' % r)
        v_fecha.append('C%d' % r)
        motor.f(ws, 'G%d' % r,
                '=IFERROR(IF($C{r}="","",DATE(YEAR($C{r}),MONTH($C{r}),1)),"")'
                .format(r=r), fmt=MES)

    motor.dv_lista(ws, v_plat, PLATAFORMAS, titulo='Plataforma no válida')
    motor.dv_lista(ws, v_tema, TEMAS, titulo='Tema no válido')
    motor.dv_lista(ws, v_resp, SI_NO, titulo='Marca Sí o No')
    motor.dv_numerica(ws, v_est, minimo=1, maximo=5, titulo='Estrellas',
                      mensaje='Escribe las estrellas de la reseña, de 1 a 5.')
    motor.dv_fecha(ws, v_fecha)
    motor.semaforo_texto(ws, 'F%d:F%d' % (S_INI, S_FIN), (
        ('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('No', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
    motor.semaforo_isnumber(ws, 'D%d:D%d' % (S_INI, S_FIN), '$D%d' % S_INI,
                            '<=', '2')
    motor.val(ws, 'A%d' % (S_FIN + 2),
              'Responder una reseña mala no es discutir: es dejar escrito, '
              'para el que la lee después, qué pasó y qué has cambiado.',
              wrap=True)
    setup(ws, titulos='$%d:$%d' % (S_CAB, S_CAB))
    return ws


# --------------------------------------------------------------------------
# Hoja «Resumen»
# --------------------------------------------------------------------------
Q = "'Registro de Quejas'"
RC = "'Reclamaciones Formales'"
RS = "'Reseñas'"


def hoja_resumen(wb):
    ws = wb.create_sheet('Resumen')
    cabecera(ws, 'Resumen — dónde se repite el problema y si llegas a tiempo')
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 60

    seccion(ws, 'A4', 'OBJETIVOS DE LA CASA — celdas verdes')
    etiqueta_valor(ws, X_OBJ_SLA, 'Objetivo de SLA cumplido (%)',
                   'D%d' % X_OBJ_SLA, fmt=PCT, valor=0.90, verde_=True,
                   nota='Por debajo de este porcentaje, la celda de '
                        'cumplimiento se pone en rojo.', col_nota='H')
    etiqueta_valor(ws, X_OBJ_RESP, 'Objetivo de reseñas respondidas (%)',
                   'D%d' % X_OBJ_RESP, fmt=PCT, valor=0.90, verde_=True,
                   nota='Responder también las buenas: es lo que sostiene la '
                        'media.', col_nota='H')
    etiqueta_valor(ws, X_OBJ_ESTRELLAS, 'Objetivo de media de estrellas',
                   'D%d' % X_OBJ_ESTRELLAS, fmt=DEC2, valor=4.30, verde_=True,
                   nota='Objetivo de la casa, no un dato de mercado.',
                   col_nota='H')
    etiqueta_valor(ws, X_TOLERANCIA,
                   'Tolerancia para leer la variación entre meses (estrellas)',
                   'D%d' % X_TOLERANCIA, fmt=DEC2, valor=0.05, verde_=True,
                   nota='Por debajo de esta diferencia, el mes se lee como '
                        '«Igual»: no toda oscilación es una tendencia.',
                   col_nota='H')
    motor.dv_porcentaje(ws, ['D%d' % X_OBJ_SLA, 'D%d' % X_OBJ_RESP],
                        titulo='Objetivo (%)',
                        prompt='Se escribe en tanto por uno: 0,90 = 90 %.')
    motor.dv_numerica(ws, ['D%d' % X_OBJ_ESTRELLAS], minimo=1, maximo=5,
                      titulo='Media de estrellas',
                      mensaje='Escribe la media objetivo, de 1 a 5.')
    motor.dv_numerica(ws, ['D%d' % X_TOLERANCIA], minimo=0, maximo=5,
                      titulo='Tolerancia',
                      mensaje='Escribe la diferencia mínima en estrellas para '
                              'considerar que un mes mejora o empeora.')

    seccion(ws, 'A10', 'QUEJAS — lo calcula el libro')
    r = X_Q_INI
    etiqueta_valor(ws, r, 'Quejas registradas', 'D%d' % r, fmt=ENT,
                   formula='=COUNTIF(%s!$B$%d:$B$%d,"<>")' % (Q, Q_INI, Q_FIN))
    etiqueta_valor(
        ws, r + 1, 'Motivo más repetido', 'D%d' % (r + 1), destacar=True,
        formula='=IFERROR(INDEX($A${a}:$A${b},MATCH(MAX($C${a}:$C${b}),'
                '$C${a}:$C${b},0)),"")'.format(a=X_MOT_INI, b=X_MOT_FIN))
    etiqueta_valor(
        ws, r + 2, 'Quejas de ese motivo', 'D%d' % (r + 2), fmt=ENT,
        formula='=IFERROR(MAX($C${a}:$C${b}),"")'.format(a=X_MOT_INI,
                                                        b=X_MOT_FIN))
    etiqueta_valor(
        ws, r + 3, 'Gravedad más repetida', 'D%d' % (r + 3), fmt=ENT,
        formula='=IFERROR(INDEX($A${a}:$A${b},MATCH(MAX($C${a}:$C${b}),'
                '$C${a}:$C${b},0)),"")'.format(a=X_GRA_INI, b=X_GRA_FIN))
    etiqueta_valor(
        ws, r + 4, 'Quejas de esa gravedad', 'D%d' % (r + 4), fmt=ENT,
        formula='=IFERROR(MAX($C${a}:$C${b}),"")'.format(a=X_GRA_INI,
                                                        b=X_GRA_FIN))
    etiqueta_valor(
        ws, r + 5, 'Tiempo medio hasta el cierre (horas)', 'D%d' % (r + 5),
        fmt=DEC1,
        formula='=IFERROR(AVERAGEIFS({q}!$I${a}:$I${b},{q}!$I${a}:$I${b},'
                '">=0"),"")'.format(q=Q, a=Q_INI, b=Q_FIN))
    etiqueta_valor(
        ws, r + 6, 'Quejas cerradas dentro del SLA', 'D%d' % (r + 6), fmt=ENT,
        formula='=COUNTIF({q}!$K${a}:$K${b},"Sí")'.format(q=Q, a=Q_INI,
                                                         b=Q_FIN))
    etiqueta_valor(
        ws, r + 7, 'Quejas cerradas fuera del SLA', 'D%d' % (r + 7), fmt=ENT,
        formula='=COUNTIF({q}!$K${a}:$K${b},"No")'.format(q=Q, a=Q_INI,
                                                         b=Q_FIN))
    etiqueta_valor(
        ws, r + 8, 'SLA CUMPLIDO (%)', 'D%d' % (r + 8), fmt=PCT,
        destacar=True,
        formula='=IFERROR(IF($D${a}+$D${b}=0,"",$D${a}/($D${a}+$D${b})),"")'
                .format(a=r + 6, b=r + 7))
    motor.semaforo_isnumber(ws, 'D%d' % (r + 8), '$D$%d' % (r + 8), '<',
                            '$D$%d' % X_OBJ_SLA)

    seccion(ws, 'A%d' % (X_MOT_CAB - 1), 'QUEJAS POR MOTIVO — el motivo que se '
                                         'repite es un problema de proceso')
    encabezados(ws, X_MOT_CAB, [
        ('A', 'Motivo', None), ('B', '', None), ('C', 'Quejas', None),
        ('D', 'Porcentaje del total', None),
        ('E', 'Horas medias hasta el cierre', None),
        ('F', 'Fuera del SLA', None), ('G', '', None), ('H', '', None),
    ], alto=34)
    for i, m in enumerate(DE.MOTIVOS_QUEJA):
        rr = X_MOT_INI + i
        motor.val(ws, 'A%d' % rr, m)
        motor.f(ws, 'C%d' % rr,
                '=COUNTIF({q}!$D${a}:$D${b},$A{r})'.format(q=Q, a=Q_INI,
                                                           b=Q_FIN, r=rr),
                fmt=ENT)
        motor.f(ws, 'D%d' % rr,
                '=IFERROR(IF($D${t}=0,"",$C{r}/$D${t}),"")'
                .format(t=X_Q_INI, r=rr), fmt=PCT)
        motor.f(ws, 'E%d' % rr,
                '=IFERROR(AVERAGEIFS({q}!$I${a}:$I${b},{q}!$D${a}:$D${b},'
                '$A{r},{q}!$I${a}:$I${b},">=0"),"")'.format(q=Q, a=Q_INI,
                                                           b=Q_FIN, r=rr),
                fmt=DEC1)
        motor.f(ws, 'F%d' % rr,
                '=COUNTIFS({q}!$D${a}:$D${b},$A{r},{q}!$K${a}:$K${b},"No")'
                .format(q=Q, a=Q_INI, b=Q_FIN, r=rr), fmt=ENT)
    motor.regla_expresion(
        ws, 'C%d:C%d' % (X_MOT_INI, X_MOT_FIN),
        '=AND(ISNUMBER($C{a}),$C{a}=MAX($C${a}:$C${b}),$C{a}>0)'
        .format(a=X_MOT_INI, b=X_MOT_FIN),
        bg=motor.CF_AMBAR_BG, fg=motor.CF_AMBAR_FG)

    seccion(ws, 'A%d' % (X_GRA_CAB - 1), 'QUEJAS POR GRAVEDAD')
    encabezados(ws, X_GRA_CAB, [
        ('A', 'Gravedad', None), ('B', 'Qué significa', None),
        ('C', 'Quejas', None), ('D', 'SLA (horas)', None),
        ('E', 'Horas medias hasta el cierre', None),
        ('F', 'Dentro del SLA', None), ('G', 'Cumplimiento (%)', None),
        ('H', '', None),
    ], alto=34)
    for i, (g, texto) in enumerate(DE.PARAMETROS_QUEJAS['escala_gravedad']):
        rr = X_GRA_INI + i
        motor.val(ws, 'A%d' % rr, g, fmt=ENT)
        motor.val(ws, 'B%d' % rr, texto, wrap=True)
        motor.f(ws, 'C%d' % rr,
                '=COUNTIF({q}!$E${a}:$E${b},$A{r})'.format(q=Q, a=Q_INI,
                                                           b=Q_FIN, r=rr),
                fmt=ENT)
        motor.f(ws, 'D%d' % rr,
                '=IFERROR(IF($A{r}="","",INDEX(Parámetros!$C${a}:$C${b},'
                'MATCH($A{r},Parámetros!$A${a}:$A${b},0))),"")'
                .format(r=rr, a=P_SLA_INI, b=P_SLA_FIN), fmt=ENT)
        motor.f(ws, 'E%d' % rr,
                '=IFERROR(AVERAGEIFS({q}!$I${a}:$I${b},{q}!$E${a}:$E${b},'
                '$A{r},{q}!$I${a}:$I${b},">=0"),"")'.format(q=Q, a=Q_INI,
                                                           b=Q_FIN, r=rr),
                fmt=DEC1)
        motor.f(ws, 'F%d' % rr,
                '=COUNTIFS({q}!$E${a}:$E${b},$A{r},{q}!$K${a}:$K${b},"Sí")'
                .format(q=Q, a=Q_INI, b=Q_FIN, r=rr), fmt=ENT)
        motor.f(ws, 'G%d' % rr,
                '=IFERROR(IF($C{r}=0,"",$F{r}/$C{r}),"")'.format(r=rr),
                fmt=PCT)
    motor.semaforo_isnumber(
        ws, 'G%d:G%d' % (X_GRA_INI, X_GRA_FIN), '$G%d' % X_GRA_INI, '<',
        '$D$%d' % X_OBJ_SLA)

    seccion(ws, 'A%d' % (X_REC_INI - 1),
            'RECLAMACIONES FORMALES — el plazo lo fija tu comunidad')
    etiqueta_valor(
        ws, X_REC_INI, 'Reclamaciones registradas', 'D%d' % X_REC_INI,
        fmt=ENT,
        formula='=COUNTIF({rc}!$B${a}:$B${b},"<>")'.format(rc=RC, a=R_INI,
                                                          b=R_FIN))
    etiqueta_valor(
        ws, X_REC_INI + 1, 'Contestadas dentro de plazo',
        'D%d' % (X_REC_INI + 1), fmt=ENT,
        formula='=COUNTIF({rc}!$I${a}:$I${b},"Sí")'.format(rc=RC, a=R_INI,
                                                          b=R_FIN))
    etiqueta_valor(
        ws, X_REC_INI + 2, 'CONTESTADAS FUERA DE PLAZO',
        'D%d' % (X_REC_INI + 2), fmt=ENT, destacar=True,
        formula='=COUNTIF({rc}!$I${a}:$I${b},"No")'.format(rc=RC, a=R_INI,
                                                          b=R_FIN))
    etiqueta_valor(
        ws, X_REC_INI + 3, 'Contestadas dentro de plazo (%)',
        'D%d' % (X_REC_INI + 3), fmt=PCT,
        formula='=IFERROR(IF($D${a}+$D${b}=0,"",$D${a}/($D${a}+$D${b})),"")'
                .format(a=X_REC_INI + 1, b=X_REC_INI + 2))
    motor.regla_expresion(
        ws, 'D%d' % (X_REC_INI + 2),
        '=AND(ISNUMBER($D${r}),$D${r}>0)'.format(r=X_REC_INI + 2))
    motor.val(ws, 'A%d' % (X_REC_INI + 4),
              'Una reclamación fuera de plazo no se arregla contestando '
              'después: se evita poniendo la fecha límite en el calendario el '
              'mismo día en que entregas la hoja.', wrap=True)

    seccion(ws, 'A%d' % (X_RES_INI - 1), 'RESEÑAS')
    etiqueta_valor(
        ws, X_RES_INI, 'Reseñas registradas', 'D%d' % X_RES_INI, fmt=ENT,
        formula='=COUNTIF({rs}!$C${a}:$C${b},"<>")'.format(rs=RS, a=S_INI,
                                                          b=S_FIN))
    etiqueta_valor(
        ws, X_RES_INI + 1, 'Media de estrellas', 'D%d' % (X_RES_INI + 1),
        fmt=DEC2, destacar=True,
        formula='=IFERROR(AVERAGEIFS({rs}!$D${a}:$D${b},{rs}!$D${a}:$D${b},'
                '">=0"),"")'.format(rs=RS, a=S_INI, b=S_FIN))
    etiqueta_valor(
        ws, X_RES_INI + 2, 'Reseñas respondidas', 'D%d' % (X_RES_INI + 2),
        fmt=ENT,
        formula='=COUNTIF({rs}!$F${a}:$F${b},"Sí")'.format(rs=RS, a=S_INI,
                                                          b=S_FIN))
    etiqueta_valor(
        ws, X_RES_INI + 3, 'Reseñas respondidas (%)', 'D%d' % (X_RES_INI + 3),
        fmt=PCT,
        formula='=IFERROR(IF($D${t}=0,"",$D${r}/$D${t}),"")'
                .format(t=X_RES_INI, r=X_RES_INI + 2))
    motor.semaforo_isnumber(ws, 'D%d' % (X_RES_INI + 1),
                            '$D$%d' % (X_RES_INI + 1), '<',
                            '$D$%d' % X_OBJ_ESTRELLAS)
    motor.semaforo_isnumber(ws, 'D%d' % (X_RES_INI + 3),
                            '$D$%d' % (X_RES_INI + 3), '<',
                            '$D$%d' % X_OBJ_RESP)

    seccion(ws, 'A%d' % (X_MES_CAB - 1),
            'RESEÑAS POR MES — aquí se ve la caída y la recuperación')
    encabezados(ws, X_MES_CAB, [
        ('A', 'Primer día del mes', None), ('B', '', None),
        ('C', 'Reseñas', None), ('D', 'Media de estrellas', None),
        ('E', 'Respondidas', None), ('F', 'Respondidas (%)', None),
        ('G', 'Lectura frente al mes anterior', None), ('H', '', None),
    ], alto=34)
    meses_sembrados = ['2026-03-01', '2026-04-01', '2026-05-01', '2026-06-01',
                       '2026-07-01', '2026-08-01']
    for i in range(X_MES_FIN - X_MES_INI + 1):
        rr = X_MES_INI + i
        if i < len(meses_sembrados):
            motor.val(ws, 'A%d' % rr, DE._fecha(meses_sembrados[i]), fmt=MES)
        else:
            ws['A%d' % rr].number_format = MES
        motor.verde(ws, 'A%d' % rr)
        sig = ('DATE(YEAR($A{r}),MONTH($A{r})+1,1)').format(r=rr)
        motor.f(ws, 'C%d' % rr,
                '=IF($A{r}="","",IFERROR(COUNTIFS({rs}!$C${a}:$C${b},">="&$A{r},'
                '{rs}!$C${a}:$C${b},"<"&{s}),""))'
                .format(r=rr, rs=RS, a=S_INI, b=S_FIN, s=sig), fmt=ENT)
        motor.f(ws, 'D%d' % rr,
                '=IF($A{r}="","",IFERROR(AVERAGEIFS({rs}!$D${a}:$D${b},'
                '{rs}!$C${a}:$C${b},">="&$A{r},{rs}!$C${a}:$C${b},"<"&{s}),""))'
                .format(r=rr, rs=RS, a=S_INI, b=S_FIN, s=sig), fmt=DEC2)
        motor.f(ws, 'E%d' % rr,
                '=IF($A{r}="","",IFERROR(COUNTIFS({rs}!$C${a}:$C${b},">="&$A{r},'
                '{rs}!$C${a}:$C${b},"<"&{s},{rs}!$F${a}:$F${b},"Sí"),""))'
                .format(r=rr, rs=RS, a=S_INI, b=S_FIN, s=sig), fmt=ENT)
        motor.f(ws, 'F%d' % rr,
                '=IFERROR(IF(OR($C{r}="",$C{r}=0),"",$E{r}/$C{r}),"")'
                .format(r=rr), fmt=PCT)
        if i == 0:
            motor.val(ws, 'G%d' % rr, 'Primera medición')
        else:
            motor.f(ws, 'G%d' % rr,
                    '=IFERROR(IF(OR($D{r}="",$D{p}=""),"",'
                    'IF($D{r}-$D{p}>$D${t},"Mejora",'
                    'IF($D{p}-$D{r}>$D${t},"Empeora","Igual"))),"")'
                    .format(r=rr, p=rr - 1, t=X_TOLERANCIA), bold=True)
    motor.semaforo_texto(ws, 'G%d:G%d' % (X_MES_INI, X_MES_FIN), (
        ('Mejora', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
        ('Empeora', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
        ('Igual', motor.CF_GRIS_BG, motor.CF_GRIS_FG)))
    motor.semaforo_isnumber(ws, 'D%d:D%d' % (X_MES_INI, X_MES_FIN),
                            '$D%d' % X_MES_INI, '<',
                            '$D$%d' % X_OBJ_ESTRELLAS)
    motor.val(ws, 'A%d' % (X_MES_FIN + 2),
              'Cruza este bloque con la tabla de motivos: si el mes en que '
              'caen las estrellas es el mismo en que se concentran las quejas '
              'por un motivo, ya sabes qué arreglar.', wrap=True)
    motor.val(ws, 'A%d' % (X_MES_FIN + 3),
              'Los meses vacíos no se cuentan: la fila se queda en blanco. '
              'Añade el primer día de cada mes nuevo en la columna verde.',
              wrap=True)
    setup(ws)
    return ws


# --------------------------------------------------------------------------
def mapa():
    r = X_Q_INI
    celdas_resumen = {
        'Objetivo de SLA cumplido (%)': 'D%d' % X_OBJ_SLA,
        'Objetivo de reseñas respondidas (%)': 'D%d' % X_OBJ_RESP,
        'Objetivo de media de estrellas': 'D%d' % X_OBJ_ESTRELLAS,
        'Tolerancia de lectura entre meses': 'D%d' % X_TOLERANCIA,
        'Quejas registradas': 'D%d' % r,
        'Motivo más repetido': 'D%d' % (r + 1),
        'Quejas del motivo más repetido': 'D%d' % (r + 2),
        'Gravedad más repetida': 'D%d' % (r + 3),
        'Quejas de la gravedad más repetida': 'D%d' % (r + 4),
        'Tiempo medio hasta el cierre (horas)': 'D%d' % (r + 5),
        'Quejas cerradas dentro del SLA': 'D%d' % (r + 6),
        'Quejas cerradas fuera del SLA': 'D%d' % (r + 7),
        'SLA cumplido (%)': 'D%d' % (r + 8),
        'Reclamaciones registradas': 'D%d' % X_REC_INI,
        'Reclamaciones contestadas dentro de plazo': 'D%d' % (X_REC_INI + 1),
        'Reclamaciones contestadas fuera de plazo': 'D%d' % (X_REC_INI + 2),
        'Reclamaciones dentro de plazo (%)': 'D%d' % (X_REC_INI + 3),
        'Reseñas registradas': 'D%d' % X_RES_INI,
        'Media de estrellas': 'D%d' % (X_RES_INI + 1),
        'Reseñas respondidas': 'D%d' % (X_RES_INI + 2),
        'Reseñas respondidas (%)': 'D%d' % (X_RES_INI + 3),
    }
    for i, m in enumerate(DE.MOTIVOS_QUEJA):
        celdas_resumen['Motivo «%s»: quejas' % m] = 'C%d' % (X_MOT_INI + i)
        celdas_resumen['Motivo «%s»: porcentaje' % m] = 'D%d' % (X_MOT_INI + i)
        celdas_resumen['Motivo «%s»: horas medias' % m] = \
            'E%d' % (X_MOT_INI + i)
    for i, (g, _t) in enumerate(DE.PARAMETROS_QUEJAS['escala_gravedad']):
        celdas_resumen['Gravedad %d: quejas' % g] = 'C%d' % (X_GRA_INI + i)
        celdas_resumen['Gravedad %d: SLA (horas)' % g] = 'D%d' % (X_GRA_INI + i)
        celdas_resumen['Gravedad %d: cumplimiento (%%)' % g] = \
            'G%d' % (X_GRA_INI + i)
    etiquetas_mes = ['marzo 2026', 'abril 2026', 'mayo 2026', 'junio 2026',
                     'julio 2026', 'agosto 2026']
    for i, et in enumerate(etiquetas_mes):
        celdas_resumen['Reseñas de %s: media de estrellas' % et] = \
            'D%d' % (X_MES_INI + i)
        celdas_resumen['Reseñas de %s: reseñas' % et] = 'C%d' % (X_MES_INI + i)
        celdas_resumen['Reseñas de %s: lectura' % et] = 'G%d' % (X_MES_INI + i)

    celdas_param = {
        'SLA de la gravedad 1 (horas)': 'C%d' % P_SLA_INI,
        'SLA de la gravedad 2 (horas)': 'C%d' % (P_SLA_INI + 1),
        'SLA de la gravedad 3 (horas)': 'C%d' % P_SLA_FIN,
        'Horas que tiene un día': 'C%d' % P_HORAS,
        'Plazo legal de Cataluña (días)': 'C%d' % P_PLZ_INI,
        'Plazo legal de Andalucía (días)': 'C%d' % (P_PLZ_INI + 1),
        # C%d de «Otras comunidades» se deja VACÍA a propósito (SPEC §2.2,
        # fila 3): no se cita como celda con dato, sólo vive en la tabla.
        'Nota legal de Cataluña': 'F%d' % P_PLZ_INI,
        'Nota legal de Andalucía': 'F%d' % (P_PLZ_INI + 1),
    }
    return {
        'fichero': NOMBRE + '.xlsx',
        'producto': PRODUCTO,
        'hojas': {
            'Parámetros': {
                'celdas': celdas_param,
                'tablas': [
                    {'titulo': 'SLA de la casa por gravedad',
                     'cols': [['Gravedad (1-3)', 'A', 'num'],
                              ['Qué significa', 'B', 'txt'],
                              ['SLA de cierre (horas)', 'C', 'num']],
                     'filas': [P_SLA_INI, P_SLA_FIN]},
                    {'titulo': 'Plazo legal de respuesta por comunidad',
                     'cols': [['Comunidad', 'A', 'txt'],
                              ['Plazo legal', 'B', 'txt'],
                              ['Días para el aviso', 'C', 'num'],
                              ['Se cuentan', 'D', 'txt'],
                              ['Norma, verificación y URL', 'F', 'txt']],
                     'filas': [P_PLZ_INI, P_PLZ_FIN]},
                    {'titulo': 'Motivos de queja (lista cerrada)',
                     'cols': [['Motivo', 'A', 'txt'],
                              ['Qué recoge', 'B', 'txt']],
                     'filas': [P_MOT_INI, P_MOT_FIN]},
                    {'titulo': 'Canales de entrada',
                     'cols': [['Canal', 'A', 'txt']],
                     'filas': [P_CAN_INI, P_CAN_FIN]},
                ],
            },
            'Registro de Quejas': {
                'celdas': {},
                'tablas': [
                    {'titulo': '30 quejas de ejemplo y 20 filas libres',
                     'cols': [['#', 'A', 'num'],
                              ['Fecha de la queja', 'B', 'txt'],
                              ['Canal', 'C', 'txt'], ['Motivo', 'D', 'txt'],
                              ['Gravedad (1-3)', 'E', 'num'],
                              ['Responsable', 'F', 'txt'],
                              ['Acción tomada', 'G', 'txt'],
                              ['Fecha de cierre', 'H', 'txt'],
                              ['Horas hasta el cierre', 'I', 'num'],
                              ['SLA de esa gravedad (horas)', 'J', 'num'],
                              ['SLA cumplido', 'K', 'txt'],
                              ['Notas', 'L', 'txt']],
                     'filas': [Q_INI, Q_FIN]},
                ],
            },
            'Reclamaciones Formales': {
                'celdas': {
                    'Días transcurridos de la reclamación 2': 'F%d' % (R_INI + 1),
                    'Plazo aplicable a la reclamación 2': 'G%d' % (R_INI + 1),
                    'Dentro de plazo de la reclamación 2': 'I%d' % (R_INI + 1),
                },
                'tablas': [
                    {'titulo': '3 reclamaciones de ejemplo y 17 filas libres',
                     'cols': [['#', 'A', 'num'],
                              ['Fecha de entrega de la hoja', 'B', 'txt'],
                              ['Número de hoja', 'C', 'txt'],
                              ['Comunidad', 'D', 'txt'],
                              ['Fecha de respuesta por escrito', 'E', 'txt'],
                              ['Días naturales transcurridos', 'F', 'num'],
                              ['Plazo de la comunidad (días)', 'G', 'num'],
                              ['Se cuentan', 'H', 'txt'],
                              ['Dentro de plazo', 'I', 'txt'],
                              ['Objeto de la reclamación y respuesta', 'J',
                               'txt']],
                     'filas': [R_INI, R_FIN]},
                ],
            },
            'Reseñas': {
                'celdas': {},
                'tablas': [
                    {'titulo': '40 reseñas de ejemplo y 20 filas libres',
                     'cols': [['#', 'A', 'num'], ['Plataforma', 'B', 'txt'],
                              ['Fecha', 'C', 'txt'],
                              ['Estrellas (1-5)', 'D', 'num'],
                              ['Tema', 'E', 'txt'],
                              ['Respondida', 'F', 'txt'], ['Mes', 'G', 'txt'],
                              ['Notas', 'H', 'txt']],
                     'filas': [S_INI, S_FIN]},
                ],
            },
            'Resumen': {
                'celdas': celdas_resumen,
                'tablas': [
                    {'titulo': 'Quejas por motivo',
                     'cols': [['Motivo', 'A', 'txt'], ['Quejas', 'C', 'num'],
                              ['Porcentaje del total', 'D', 'pct1'],
                              ['Horas medias hasta el cierre', 'E', 'num'],
                              ['Fuera del SLA', 'F', 'num']],
                     'filas': [X_MOT_INI, X_MOT_FIN]},
                    {'titulo': 'Quejas por gravedad',
                     'cols': [['Gravedad', 'A', 'num'],
                              ['Qué significa', 'B', 'txt'],
                              ['Quejas', 'C', 'num'],
                              ['SLA (horas)', 'D', 'num'],
                              ['Horas medias hasta el cierre', 'E', 'num'],
                              ['Dentro del SLA', 'F', 'num'],
                              ['Cumplimiento (%)', 'G', 'pct1']],
                     'filas': [X_GRA_INI, X_GRA_FIN]},
                    {'titulo': 'Reseñas por mes',
                     'cols': [['Primer día del mes', 'A', 'txt'],
                              ['Reseñas', 'C', 'num'],
                              ['Media de estrellas', 'D', 'num'],
                              ['Respondidas', 'E', 'num'],
                              ['Respondidas (%)', 'F', 'pct1'],
                              ['Lectura frente al mes anterior', 'G', 'txt']],
                     'filas': [X_MES_INI, X_MES_FIN]},
                ],
            },
        },
    }


# --------------------------------------------------------------------------
def main():
    destino = os.path.join(AQUI, 'build')
    if not os.path.isdir(destino):
        os.makedirs(destino)
    ruta = os.path.join(destino, NOMBRE + '.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_instrucciones(wb)
    hoja_parametros(wb)
    hoja_quejas(wb)
    hoja_reclamaciones(wb)
    hoja_resenas(wb)
    hoja_resumen(wb)

    verdes = {}
    for ws in wb.worksheets:
        motor.retirar_verde_de_calculadas(ws)
        verdes[ws.title] = motor.proteger(ws)

    wb.properties.creator = 'AI Chef Pro'
    wb.properties.lastModifiedBy = 'AI Chef Pro'
    wb.properties.title = TITULO_LIBRO
    wb.properties.subject = SUBJECT
    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)

    with open(os.path.join(destino, 'mapa-' + NOMBRE + '.json'), 'w',
              encoding='utf-8') as fh:
        json.dump(mapa(), fh, ensure_ascii=False, indent=1)

    print('OK', ruta)
    print('formulas registradas:', len(motor.REGISTRO))
    for hoja, n in verdes.items():
        print('  verdes %-24s %d' % (hoja, n))


if __name__ == '__main__':
    main()
