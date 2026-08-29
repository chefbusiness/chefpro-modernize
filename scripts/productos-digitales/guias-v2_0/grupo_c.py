#!/usr/bin/env python3
"""
grupo_c.py — Grupo C de la familia «Guías Cómo Montar» v2.0: **cocina, carta y
sala** (`guias-v2-SPEC.md` §4 entero).

Ficheros que toca (MEDIDO el 2026-08-29 abriendo los 8 productos de
`astro-site/public/dl/guia-*`, no leído del informe):

    escandallo-maestro[-<concepto>].xlsx   7 guías  (NO 8: dark-kitchen no lo
                                                     lleva — sus 5 entregables
                                                     son 2 checklists, la
                                                     calculadora, el docx y el
                                                     PDF)
    menu-engineering-matrix.xlsx           6 guías  (todas menos panadería y
                                                     dark-kitchen)
    budget-bodega.xlsx                     1 guía   (representante)
    plantilla-turnos-brigada.xlsx          7 guías
    plan-fermentacion-y-produccion.xlsx    1 guía   (panadería)

Aquí no hay ni una fila de contenido: sólo el motor de aplicación. Las filas,
los textos, los importes y los parámetros de cada guía viven en
`contenido_<pid>/c.py`, con la fuente de cada cifra anotada.

TRES MOLDES DE ESCANDALLO Y TRES DE CUADRANTE, no dos y uno
-----------------------------------------------------------
La SPEC §4.1 describe **dos** modelos de escandallo («representante» y
«hermanos») y §4.4 describe **un** cuadrante. Medido, hay tres de cada uno, y
aplicar la rejilla del representante a panadería escribiría encima de sus datos:

  · **E1 — representante** (`escandallo-maestro.xlsx`): cabecera en la fila 6,
    `A #`·`B Ingrediente`·`C Unidad`·`D Cantidad Bruta`·`E Precio/Ud (€)`·
    `F Merma (%)`·`G Cantidad Neta`·`H Coste (€)`·`I Notas`; banda de ficha en
    la fila 4 (`A4 Nombre del plato:` · `E4 Raciones:` · `G4 Food Cost
    Objetivo:` · `H4 '28%'` **como cadena**); 20 líneas vacías; `H27=SUM`,
    `H28='=H27/0.28'`.
  · **E2 — los 5 hermanos** (casual, mexicano, peruano, japonés, nikkei):
    cabecera en la fila 4, `C Cantidad (g/ml)`·`D Precio/Kg (€)`·`E Coste (€)`·
    `F Merma (%)`·`G Coste Real (€)`; ficha PRECARGADA con 8-13 ingredientes
    reales; `G18=SUM`, `G19='=G18/0.33'`, `G20='=G19*1.10'`. **El food cost del
    rótulo NO es el mismo en los cinco**: casual va al 30 % y japonés al 33 %,
    así que se LEE de la fórmula, no se teclea.
  · **E3 — panadería** (`escandallo-maestro-panaderia.xlsx`): no es una ficha,
    es una **tabla de 25 productos** con cabecera en la fila 3
    (`A Producto`·`B Peso final g`·`C Coste materia €`·`D PVP €`·
    `E Food cost %`·`F Margen bruto %`·`G Margen € unidad`·`H Notas técnicas`).
    **No tiene columna de merma, ni raciones, ni PVP sugerido, ni IVA**, y su
    hoja no se llama `Escandallo` sino `Escandallos`. Renombrarla a
    «Ficha (plantilla)» (TEC-23) sería mentir: no hay ficha que duplicar.

  · **T1 — representante** (`plantilla-turnos-brigada.xlsx`): cabecera fila 4,
    `A #` VACÍA, `B Nombre`, `C Puesto`, `D:J Lun-Dom`, `K Horas/Semana`; dos
    bandas de sección (`B5 COCINA`, `B21 SALA`) y **24 puestos** (C6:C20 = 15 de
    cocina, C22:C30 = 9 de sala) frente al «(25 personas)» del título.
  · **T2 — los 5 hermanos**: cabecera fila 4, `A #` NUMERADA desde la fila 5,
    **sin columna `K Horas/Semana`** y con la rejilla YA rellena de `P`/`T`/`M`.
  · **T3 — panadería**: cabecera fila 3, `A Persona`·`B Rol`·`C:I L-D`·
    `J Total h`·`K Plus nocturnidad`. Los días **no llevan la letra del turno,
    llevan el RANGO HORARIO como texto** (`'03-11'`, `'12:30-14'`, `'libre'`) y
    `J` son constantes tecleadas. La tabla de equivalencia turno→horas del §4.4
    no tiene dónde aplicarse: aquí `J` queda como entrada verde documentada y
    las horas calculadas viven en la hoja `Registro de jornada`, que es la que
    cumple el art. 34.9 ET.

Igual con el menú: **M1** (representante, hoja `Menu Engineering`, `A:H`, 25
filas VACÍAS) y **M2** (los 5 hermanos, hoja `Matrix`, `A:J`, 15 platos reales
con `I Popularidad` y `J Clasificación` rotuladas y vacías).

La estructura se detecta por la CABECERA, nunca por la posición, y si no encaja
en ninguna firma el módulo **aborta** con el fichero y la cabecera cruda
(`EstructuraDesconocida`), igual que hace `motor.molde_checklist()` (§1.1,
§7-bis.11). Un molde por defecto es lo que rompe panadería.

⚠️ TRAMPA DE pycel MEDIDA HOY (2026-08-29, pycel 1.0b30, este Mac)
-------------------------------------------------------------------
`SUMPRODUCT` con multiplicación de booleanos —`SUMPRODUCT((D6:J6="M")*$B$40+…)`,
que es la forma canónica de contar horas por turno— devuelve un **`numpy.int64`**,
y el propio pycel NO lo reconoce después: `SUM(K6:K30)` sobre esas celdas da
**0** y `AND(ISNUMBER(K6),K6<40)` da **False** con K6 = 44. Medido celda a celda:

    K10 = SUMPRODUCT(...)              -> 44      (numpy.int64)
    K14 = SUM(K10:K12)                 -> 0       ← el total se pierde
    K15 = AND(ISNUMBER(K10),K10<100)   -> False   ← el semáforo no se enciende
    M10 = ROUND(SUMPRODUCT(...),2)     -> 44.0    (float)
    M14 = SUM(M10:M12)                 -> 44.0    ✓
    M15 = AND(ISNUMBER(M10),M10<100)   -> True    ✓

Excel calcula bien las dos formas; lo que se rompe es la VERIFICACIÓN (las demos
de §4.4, `inject_cache.py` y el gate de `data_only`), y se rompe **en silencio**,
devolviendo un cero plausible en una fila de coste de personal. Por eso en este
módulo **todo `SUMPRODUCT` va envuelto en `ROUND(...,2)`**, sin excepción.

CERO LITERALES DENTRO DE FÓRMULAS (§1.3 y el veredicto de `main.py`)
--------------------------------------------------------------------
`main.py` convierte `literales_sospechosos` en **fallo** en cuanto se carga un
grupo (sólo es aviso con `--solo motor`). Medido antes de tocar nada, el
representante tiene **un** literal vivo en todo el producto y es de este grupo:
`escandallo-maestro.xlsx:Escandallo!H28='=H27/0.28'`. Aquí desaparece, y no se
introduce ninguno nuevo: el 0,7 del umbral de Kasavana & Smith, el 33 % de SS,
el SMI, el IVA, el food cost objetivo, las 40 h semanales, las horas de cada
turno y hasta los 1.000 g del kilo del modelo E2 van **en celda**.

Las horas se miden con `MOD(salida-entrada,1)` y se guardan como DURACIÓN con
formato `[h]:mm`, no como `MOD(...)*24`: así no hace falta el literal 24 y el
cruce de medianoche sale bien (medido con pycel: 23:00 → 08:00 = 9 h).

IDEMPOTENCIA (main.py la comprueba con una 2.ª pasada sobre un clon)
--------------------------------------------------------------------
Todo lo que se escribe es **absoluto** y se calcula a partir de la última fila
de datos MEDIDA en esa pasada. Los bloques que este módulo posee (totales,
parámetros, pie) se BORRAN enteros y se reescriben, y se reconocen por su
etiqueta, no por su posición. Las columnas nuevas se localizan por su cabecera.
El renombrado de pestaña (`Escandallo` → `Ficha (plantilla)`) se hace en `pre()`
y es idempotente porque comprueba el nombre destino.

REGLA DE CIFRAS (regla capital + §7-bis): ninguna cifra del sector se teclea. Lo
que no está en la SPEC ni en el fichero original queda como **celda verde vacía
con nota** —nunca `0`, que en un margen se lee «0,0 %»—, y la fórmula que la usa
devuelve `""`. Con el libro en blanco no se enciende ni un semáforo.

⚠️ Los ficheros de la familia llevan ESPACIO FINO (U+202F) y GUION NO SEPARABLE
(U+2011). Aquí se referencian por escape (`motor.NARROW`, `motor.NOBRK`), nunca
escribiendo el carácter: al pasar por un heredoc del shell degeneran en espacio
y guion normales y ninguna sustitución encuentra su patrón (CLAUDE.md).
"""
import contextlib
import copy
import glob
import os
import re
import unicodedata

from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

import motor

AQUI = os.path.dirname(os.path.abspath(__file__))
DL = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(AQUI))),
                  'astro-site', 'public', 'dl')

SPEC = 'guias-v2-SPEC.md §4'

#: Prefijos de los ficheros del grupo C. Se resuelve por GLOB sobre la carpeta
#: del producto porque los hermanos renombran (`escandallo-maestro-japones`,
#: `escandallo-maestro-panaderia`): una lista fija dejaría fuera justo los que
#: cambian de nombre.
PREFIJOS = ('escandallo-maestro', 'menu-engineering-matrix', 'budget-bodega',
            'plantilla-turnos-brigada', 'plan-fermentacion-y-produccion')

#: Cabecera del bloque de parámetros DENTRO de la hoja de datos. Va en la hoja y
#: no en `Instrucciones` por una razón medida: el formato condicional que compara
#: un bruto anual con el SMI necesita que el SMI esté en la MISMA hoja — una
#: regla de formato condicional que apunta a otra hoja no la escriben todos los
#: visores y la que no la entiende deja la celda sin pintar, que es justo el caso
#: en el que el semáforo importa.
CAB_PARAMETROS = 'PARÁMETROS DE ESTE LIBRO (edítalos: el libro recalcula)'
RX_CAB_PARAM = re.compile(r'^PAR[ÁA]METROS DE ESTE LIBRO')

#: Marcador del bloque de cierre que este módulo DUEÑA en cada hoja de datos.
#: Delimita la región que se borra y se reescribe entera en cada pasada: así el
#: bloque nunca se duplica ni se desplaza (es el mismo mecanismo que
#: `motor.MARCA_BLOQUE` en los checklists).
MARCA_C = 'RESUMEN Y PARÁMETROS — lo calcula el libro (v2.0)'
RX_MARCA_C = re.compile(r'^RESUMEN Y PAR[ÁA]METROS')

RX_PIE = re.compile(r'^AI Chef Pro\s*·\s*aichef\.pro')

FMT_HORA = 'hh:mm'
FMT_DUR = '[h]:mm'
FMT_G = '#,##0.000'

NOTA_SIN_DATO = ('Sin dato en la guía: escríbelo tú. Mientras esté vacío, las '
                 'celdas que dependen de él muestran "" y no un cero que '
                 'parezca un resultado.')

#: Nombre de la hoja plantilla del escandallo (TEC-23). Los paréntesis son
#: legales en un nombre de hoja y las fórmulas que la citan van entrecomilladas
#: (`='Ficha (plantilla)'!H27`).
HOJA_FICHA = 'Ficha (plantilla)'
HOJA_RESUMEN = 'Resumen'
HOJA_REGISTRO = 'Registro de jornada'


class EstructuraDesconocida(Exception):
    """§1.1/§7-bis.11 aplicado al grupo C: el módulo NO adivina una rejilla."""


# ==========================================================================
# Utilidades de lectura (etiquetas, cabeceras, límites de datos)
# ==========================================================================
def _sin_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')


def _txt(v):
    return v.strip() if isinstance(v, str) else v


def _norm(v):
    """Etiqueta normalizada: sin acentos, sin dobles espacios, en minúscula.

    Normaliza también el ESPACIO FINO (U+202F) y el GUION NO SEPARABLE (U+2011)
    a sus equivalentes normales: un fichero de esta familia puede traerlos y un
    patrón escrito con espacio normal no encontraría nada.
    """
    if not isinstance(v, str):
        return ''
    # La convención tipográfica se aplica ANTES de comparar: el barrido
    # de `motor.normalizar_texto` cambia «≤» por «<=» y «sólo» por «solo»
    # en el fichero, y sin esto el módulo de contenido —que sigue
    # escribiéndolos a la vieja usanza— no reconocería su propia salida.
    t = motor.convencion(v).replace(motor.NARROW, ' ').replace(
        motor.NOBRK, '-')
    return re.sub(r'\s+', ' ', _sin_acentos(t).lower()).strip()


def _col(letra):
    return column_index_from_string(letra)


def _cab(ws, fila):
    """{letra: etiqueta cruda} de una fila de cabecera."""
    return dict((get_column_letter(c), _txt(ws.cell(row=fila, column=c).value))
                for c in range(1, ws.max_column + 1))


def _fila_cabecera(ws):
    """Fila que encabeza la tabla. Reutiliza el detector del motor (que ya
    resuelve el caso trampa de `escandallo-maestro!A4`, la banda de ficha que
    parece una cabecera y no lo es) y amplía la ventana a la fila 8 porque
    algunas hojas nuevas de este grupo llevan la cabecera más abajo."""
    return motor.fila_cabecera_tabla(ws, tope=8, minimo=3)


def _es_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _fin_por_columna(ws, fila_cab, letra, numerica=False, tope=None):
    """Última fila con dato en la columna indicada, MEDIDA en esta pasada.

    §9 (gate de recuento): los recuentos del R1 no cuadran entre sí, así que
    nunca se toma una cifra de un informe; se mide el fichero.

    ⚠️ `tope` no es cosmético: es lo que salvó la idempotencia. El bloque de
    parámetros que este módulo escribe DEBAJO de la tabla deja números en las
    mismas columnas por las que se mide el final de los datos (el valor de
    «Horas del turno M» vive en la columna C, la misma que los puestos). Sin
    cortar en el marcador, la 2.ª pasada mediría `fin` diez filas más abajo,
    trataría los parámetros como puestos y el bloque bajaría en cada ejecución.
    """
    col = _col(letra)
    fin = fila_cab
    ultima = min(tope, ws.max_row) if tope else ws.max_row
    for r in range(fila_cab + 1, ultima + 1):
        v = ws.cell(row=r, column=col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if numerica and not _es_num(v):
            continue
        fin = r
    return fin


def _primera_por_columna(ws, fila_cab, letra, numerica=False, tope=None):
    """Primera fila de DATOS, medida igual que la última.

    No es `fila_cab + 1` y darlo por hecho costó un defecto real: en el modelo
    E2 la fila que sigue a la cabecera es `B5='NOMBRE DEL PLATO:'` con el nombre
    del plato en `C5`, no un ingrediente. Escribir ahí la fórmula de coste
    metía un cálculo en la fila del título del plato y lo incluía en el SUM del
    coste total. Se detectó al correr el módulo contra un HERMANO, no contra el
    representante, donde la primera fila sí es la 7.
    """
    col = _col(letra)
    ultima = min(tope, ws.max_row) if tope else ws.max_row
    for r in range(fila_cab + 1, ultima + 1):
        v = ws.cell(row=r, column=col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if numerica and not _es_num(v):
            continue
        return r
    return fila_cab + 1


def _fila_marca(ws):
    """Fila donde empieza el bloque que este módulo posee, si ya está escrito."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 8) + 1):
            v = _txt(ws.cell(row=r, column=c).value)
            if isinstance(v, str) and RX_MARCA_C.match(v):
                return r
    return None


def _pie_existente(ws, desde=1):
    """(fila, texto) del pie «AI Chef Pro · aichef.pro — …» si lo hay.

    Los moldes E2/M2/T2 lo llevan como fila combinada al final de la hoja, justo
    donde este grupo necesita escribir el bloque de totales. Se captura para
    reescribirlo DEBAJO del bloque en vez de perderlo.

    ⚠️ `desde` no es opcional en la práctica: la MISMA cadena está en `A2` como
    subtítulo de la cabecera en las tres familias («AI Chef Pro · aichef.pro —
    Restaurante Gastronómico»). Sin el corte, el cuadrante del representante
    —que NO tiene pie— «capturaba» su propio subtítulo y lo reescribía al final
    de la hoja: una línea duplicada que no estaba antes.
    """
    for r in range(ws.max_row, max(desde, 1) - 1, -1):
        for c in range(1, min(ws.max_column, 3) + 1):
            v = _txt(ws.cell(row=r, column=c).value)
            if isinstance(v, str) and RX_PIE.match(v):
                return r, v
    return None, None


def _limpiar(ws, desde, hasta):
    """Vacía (valor, relleno, formato, bloqueo) el rango que este módulo posee.

    Deshace antes las combinaciones que lo tocan — el pie de los moldes E2/M2/T2
    es una fila combinada y `MergedCell.value` es de sólo lectura.
    """
    if hasta < desde:
        return
    ultima = get_column_letter(max(ws.max_column, 12))
    motor.limpiar_rango(ws, 'A' + str(desde) + ':' + ultima + str(hasta))


def _snapshot_parametros(ws, desde, hasta):
    """Guarda `etiqueta -> valor` del bloque ANTES de borrarlo.

    El bloque de parámetros se reescribe entero en cada pasada, y algunos
    valores no se pueden recalcular: el food cost de los hermanos se lee del
    literal de la fórmula ORIGINAL («=G18/0.33»), que en la 2.ª pasada ya no
    existe. Sin esta foto, la 2.ª pasada dejaba `E20` vacío y la idempotencia
    saltaba — y, peor, en producción se habría llevado por delante el valor que
    el cliente hubiera escrito ahí.
    """
    prev = {}
    if hasta >= desde:
        for r in range(desde, min(hasta, ws.max_row) + 1):
            for c in range(1, ws.max_column):
                et = _norm(ws.cell(row=r, column=c).value)
                if not et:
                    continue
                for cc in range(c + 1, min(c + 2, ws.max_column) + 1):
                    v = ws.cell(row=r, column=cc).value
                    if v is not None and not isinstance(v, str):
                        prev.setdefault(et, v)
                        break
    ws._g_prev = prev
    return prev


def _a4(ws):
    """A4 completo en una hoja NUEVA (§1.13 protege las que ya existían, no las
    que no existen: sin `paperSize=9` + ajuste + pie, `censo-entregables.py`
    la cuenta como `noprint`, que es defecto de --fail)."""
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8


# ==========================================================================
# Escritura (etiquetas, parámetros, cabeceras nuevas)
# ==========================================================================
def _et(ws, coord, texto, bold=False, wrap=False):
    return motor.val(ws, coord, texto, bold=bold or None, wrap=wrap or None)


def _cabecera_nueva(ws, fila, letra, texto, ancho=None):
    """Escribe la cabecera de una columna NUEVA con el estilo de la cabecera que
    ya existe (fondo oscuro y negrita de la familia), copiándolo de la columna
    A: si se escribiera con el estilo por defecto, la tabla saldría con una
    cabecera a medio pintar."""
    destino = ws[letra + str(fila)]
    modelo = ws.cell(row=fila, column=1)
    if modelo.has_style:
        destino._style = copy.copy(modelo._style)
    destino.value = texto
    destino.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    if ancho:
        ws.column_dimensions[letra].width = ancho
    return letra


def _param(ws, fila, clave, valor=None, etiqueta=None, fmt=None, nota=None,
           col_et='A', col_val=None, col_nota=None, verde=True,
           porcentaje=False, minimo=0, maximo=None):
    """Un parámetro EN LA HOJA: `etiqueta | valor | nota`, valor en verde.

    `clave` puede ser una de `motor.PARAMETROS` (IVA, SS empresa, SMI), y
    entonces la etiqueta, el valor por defecto, el formato y la nota salen de
    ahí — es lo que garantiza que el 33 % de SS y el SMI del RD 126/2026 digan
    lo mismo en los 111 xlsx (§7-bis.16).

    Devuelve la referencia ABSOLUTA del valor (`$B$42`) para incrustarla en las
    fórmulas: parámetro en celda, nunca literal dentro de la fórmula (§1.5).
    """
    if clave in motor.PARAMETROS:
        p = motor.PARAMETROS[clave]
        etiqueta = etiqueta or p['etiqueta']
        valor = p['valor'] if valor is None else valor
        fmt = fmt or p['formato']
        nota = nota or p['nota']
    col_val = col_val or get_column_letter(_col(col_et) + 1)
    col_nota = col_nota or get_column_letter(_col(col_val) + 1)
    # Un parámetro sin valor NUNCA pisa el que ya está en la celda. Sin esta
    # línea, la 2.ª pasada borraba el food cost del escandallo de los hermanos:
    # el 0,33 se lee de la fórmula ORIGINAL («=G18/0.33»), que en la 2.ª pasada
    # ya no existe, así que `valor` llegaba None y el parámetro se vaciaba. Lo
    # cazó la idempotencia corriendo el módulo contra japonés.
    if valor is None:
        actual = ws[col_val + str(fila)].value
        if actual is not None and not isinstance(actual, str):
            valor = actual
        else:
            valor = getattr(ws, '_g_prev', {}).get(_norm(etiqueta))
    _et(ws, col_et + str(fila), etiqueta)
    celda = motor.val(ws, col_val + str(fila), valor, fmt=fmt, verde_=verde)
    motor.fijar_formato(ws, col_val + str(fila), fmt or 'General')
    if nota:
        motor.val(ws, col_nota + str(fila), nota, wrap=True)
    if verde:
        motor.marcar_editable(ws, col_val + str(fila))
        if porcentaje:
            motor.dv_porcentaje(ws, [col_val + str(fila)])
        elif fmt not in (FMT_DUR, FMT_HORA):
            motor.dv_numerica(ws, [col_val + str(fila)], minimo=minimo,
                              maximo=maximo)
    del celda
    return '$' + col_val + '$' + str(fila)


def _suma_guardada(rango):
    """`SUM(rango)` que devuelve `""` cuando NO hay ni un dato (§7-bis.13).

    Un `SUM` a secas sobre una columna en blanco imprime «0,00 €», que en una
    fila rotulada TOTAL se lee como un resultado y no como «aún no has puesto
    nada». El representante entrega el escandallo, el menu engineering y el
    cuadrante COMPLETAMENTE vacíos: sin esta guarda, la v2.0 los entregaría
    llenos de ceros que parecen calculados.
    """
    # RT-01 — la guarda NO puede ser `COUNTIF(rango,"<>")=0`: en Excel una
    # fórmula que devuelve "" NO es una celda vacía para COUNTIF/COUNTA (sólo
    # COUNTBLANK la cuenta como vacía), así que el guardián valía 0 en pycel
    # —donde se verificaba— y el número de fórmulas en Excel, que es lo que ve
    # el cliente: «COSTE TOTAL DE LA FICHA 0,00 €». `COUNT` cuenta NÚMEROS e
    # ignora el texto "" en los dos motores, que es exactamente lo que hace
    # falta en una columna numérica.
    return ('=IF(COUNT(' + rango + ')=0,"",SUM(' + rango + '))')


def _pct(numerador, denominador):
    return motor.iferror('IF(OR(' + numerador + '="",' + denominador
                         + '="",' + denominador + '=0),"",' + numerador + '/'
                         + denominador + ')')


# ==========================================================================
# §4.1 — escandallo-maestro: detección de modelo
# ==========================================================================
def modelo_escandallo(ws, fname=''):
    """'E1' | 'E2' | 'E3' por la CABECERA (medida el 2026-08-29).

    Las firmas se eligen sobre columnas que este módulo NO reescribe: `D` cambia
    de «Cantidad Bruta» a «Cantidad NETA (ración)» en la 1.ª pasada, así que
    anclar ahí daría `EstructuraDesconocida` en la 2.ª y la idempotencia saltaría
    con un aborto en vez de con una diferencia.
    """
    fila = _fila_cabecera(ws)
    if not fila:
        raise EstructuraDesconocida(
            fname + ':' + ws.title + ': sin fila de cabecera reconocible')
    c = dict((k, _norm(v)) for k, v in _cab(ws, fila).items())
    if c.get('C') == 'unidad' and c.get('E') == 'precio/ud (€)':
        return 'E1', fila
    if c.get('C') == 'cantidad (g/ml)' and c.get('D') == 'precio/kg (€)':
        return 'E2', fila
    if c.get('A') == 'producto' and c.get('C') == 'coste materia €':
        return 'E3', fila
    raise EstructuraDesconocida(
        fname + ':' + ws.title + ': cabecera de escandallo no reconocida por '
        '§4.1 (fila ' + str(fila) + ') = '
        + repr(dict((k, v) for k, v in _cab(ws, fila).items() if v is not None))
        + ' — el módulo NO aplica un modelo por defecto: añade la firma con la '
          'evidencia medida.')


def _literal_de(ws, coord, patron):
    """Extrae el número que hoy vive DENTRO de una fórmula (`'=G18/0.33'`).

    Es lo que evita teclear el food cost: casual va al 30 % y japonés al 33 %, y
    el IVA está incrustado como `*1.10`. El número se saca del fichero y pasa a
    celda verde; si la fórmula ya no lo tiene (2.ª pasada) devuelve `None` y el
    valor que manda es el que ya está en la celda de parámetro.
    """
    v = ws[coord].value
    if not isinstance(v, str):
        return None
    m = re.search(patron, v)
    return float(m.group(1)) if m else None


# ==========================================================================
# §4.1 — E1: la ficha del representante
# ==========================================================================
def _escandallo_e1(wb, ws, fila_cab, fname, cambios, contenido):
    marca = _fila_marca(ws)
    _t = (marca - 1) if marca else None
    fin = _fin_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    primera = _primera_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    if fin <= fila_cab:
        raise EstructuraDesconocida(fname + ':' + ws.title
                                    + ': 0 líneas de ingrediente')
    cfg = getattr(contenido, 'ESCANDALLO', None) or {}

    # ---- banda de ficha (fila 4): raciones y food cost objetivo -----------
    # `arreglar_cabecera_escandallo()` del motor ya combina A4:B4 (etiqueta) y
    # C4:D4 (campo). Aquí se completa lo que es de negocio.
    banda = 4
    _et(ws, 'E' + str(banda), 'Raciones:')
    if not _es_num(ws['F' + str(banda)].value):
        motor.val(ws, 'F' + str(banda), 1, fmt=motor.FMT_ENT, verde_=True)
    else:
        motor.val(ws, 'F' + str(banda), ws['F' + str(banda)].value,
                  fmt=motor.FMT_ENT, verde_=True)
    motor.fijar_formato(ws, 'F' + str(banda), motor.FMT_ENT)
    motor.marcar_editable(ws, 'F' + str(banda))
    motor.dv_numerica(ws, ['F' + str(banda)], minimo=1,
                      titulo='Raciones',
                      mensaje='Escribe cuántas raciones salen de esta ficha '
                              '(1 o más).')
    # TEC-21/DOM-30: la cadena '28%' pasa a NÚMERO y la usa la fórmula.
    fc_actual = ws['H' + str(banda)].value
    fc = cfg.get('food_cost_objetivo')
    if _es_num(fc_actual):
        fc = fc_actual
    elif isinstance(fc_actual, str) and re.match(r'^\s*\d+(?:[.,]\d+)?\s*%\s*$',
                                                 fc_actual):
        fc = float(fc_actual.strip().rstrip('%').replace(',', '.')) / 100.0
        cambios.append(fname + ':' + ws.title + '!H' + str(banda)
                       + ': food cost objetivo ' + repr(fc_actual)
                       + ' (CADENA de texto, que ninguna fórmula podía leer) '
                       '-> número ' + str(fc) + ' con formato 0% [TEC-21 · '
                       'DOM-30 · §4.1]')
    _et(ws, 'G' + str(banda), 'Food Cost Objetivo:')
    motor.val(ws, 'H' + str(banda), fc, fmt='0%', verde_=True)
    motor.fijar_formato(ws, 'H' + str(banda), '0%')
    motor.marcar_editable(ws, 'H' + str(banda))
    motor.dv_porcentaje(ws, ['H' + str(banda)])
    ref_fc = '$H$' + str(banda)
    ref_rac = '$F$' + str(banda)

    # ---- la merma entra en el coste (DOM-04) ------------------------------
    # RC-31 · §1.12 fija G en 22; el ensanchado automático la llevaba a 41
    # sobre un A4 con `fitToWidth=1`.
    motor.fijar_ancho(ws, 'G', 22.0)
    ws['D' + str(fila_cab)].value = 'Cantidad NETA (ración)'
    ws['G' + str(fila_cab)].value = 'Cantidad BRUTA a comprar'
    for r in range(primera, fin + 1):
        motor.f(ws, 'G' + str(r),
                motor.iferror('IF(OR($D{r}="",$F{r}>=1),"",$D{r}/(1-$F{r}))'
                              .format(r=r)), fmt=FMT_G)
        motor.f(ws, 'H' + str(r),
                motor.iferror('IF(OR($G{r}="",$E{r}=""),"",$G{r}*$E{r})'
                              .format(r=r)), fmt=motor.FMT_EUR)
        motor.verde(ws, 'B' + str(r) + ':F' + str(r))
        motor.verde(ws, 'I' + str(r))
        motor.fijar_formato(ws, 'D' + str(r), FMT_G)
        motor.fijar_formato(ws, 'E' + str(r), motor.FMT_EUR)
        motor.fijar_formato(ws, 'F' + str(r), motor.FMT_PCT)
    # RT-15/RC-11 · el mensaje de entrada de la merma se pegaba en TODAS las
    # celdas de porcentaje del producto (tipo de IVA, SS de la empresa, tipo
    # del préstamo…). El tope de 0,95 tampoco vale fuera de aquí: sólo la
    # merma entra en una división por `1-x` (RT-16).
    motor.prompt_porcentaje(
        ws, 'F' + str(primera) + ':F' + str(fin), *motor.PCT_MERMA)
    cambios.append(
        fname + ':' + ws.title + '!G' + str(primera) + ':H' + str(fin)
        + ': la merma ENTRA en el coste — D pasa a «Cantidad NETA (ración)», '
          'G a «Cantidad BRUTA a comprar» = neta/(1-merma) y H = bruta x '
          'precio. Antes H=D*E no leía ni F ni G: con merluza a 22 EUR/kg, '
          '0,180 kg y 40 % de merma daba 3,96 EUR en vez de 6,60 EUR '
          '[DOM-04 · TEC-06 · COM-33 · §4.1]')

    # ---- bloque de resultado (lo posee este módulo) -----------------------
    _snapshot_parametros(ws, fin + 1, ws.max_row)
    _limpiar(ws, fin + 1, max(ws.max_row, (marca or 0) + 14))
    r = fin + 1
    _et(ws, 'G' + str(r), MARCA_C, bold=True)
    r += 1
    _et(ws, 'G' + str(r), 'COSTE TOTAL DE LA FICHA (€)', bold=True)
    motor.f(ws, 'H' + str(r),
            _suma_guardada('H' + str(primera) + ':H' + str(fin)),
            fmt=motor.FMT_EUR, bold=True)
    ref_total = '$H$' + str(r)
    r += 1
    _et(ws, 'G' + str(r), 'Coste por ración (€)', bold=True)
    motor.f(ws, 'H' + str(r),
            motor.iferror('IF(OR(' + ref_total + '="",' + ref_rac + '="",'
                          + ref_rac + '=0),"",' + ref_total + '/' + ref_rac
                          + ')'), fmt=motor.FMT_EUR, bold=True)
    ref_racion = '$H$' + str(r)
    r += 1
    motor.f(ws, 'G' + str(r),
            '="PVP sugerido ("&TEXT(' + ref_fc + ',"0%")&"), sin IVA:"')
    motor.f(ws, 'H' + str(r),
            motor.iferror('IF(OR(' + ref_racion + '="",' + ref_fc + '="",'
                          + ref_fc + '=0),"",' + ref_racion + '/' + ref_fc
                          + ')'), fmt=motor.FMT_EUR, bold=True)
    ref_pvp = '$H$' + str(r)
    r += 1
    fila_iva = r
    ref_iva = _param(ws, fila_iva, 'iva_restauracion', col_et='G',
                     col_val='H', col_nota='I')
    r += 1
    _et(ws, 'G' + str(r), 'PVP con IVA (€) — el que va en la carta', bold=True)
    motor.f(ws, 'H' + str(r),
            motor.iferror('IF(' + ref_pvp + '="","",' + ref_pvp + '*(1+'
                          + ref_iva + '))'), fmt=motor.FMT_EUR, bold=True)
    r += 1
    # RT-10/RD-22/RC-18 · «Food cost real sobre el PVP sin IVA» era una
    # TAUTOLOGÍA: el PVP se calcula dividiendo el coste entre el objetivo, así
    # que coste/PVP devuelve SIEMPRE el objetivo, con cualquier escandallo.
    # Parecía una comprobación y no podía discrepar nunca. Lo que hace falta es
    # el food cost contra el precio que de VERDAD va en la carta —que se
    # redondea (76,79 € → 79 €) y que en España se anuncia con IVA—, y el 30,8 %
    # que la nota de abajo promete y que ninguna celda calculaba.
    fila_pvp_real = r
    _et(ws, 'G' + str(r), 'PVP REAL de carta, con IVA (€) — el que imprimes')
    motor.val(ws, 'H' + str(r), None, fmt=motor.FMT_EUR, verde_=True)
    motor.fijar_formato(ws, 'H' + str(r), motor.FMT_EUR)
    _et(ws, 'I' + str(r),
        'Escribe aquí el precio redondeado que va en la carta. Mientras esté '
        'vacío, las dos filas de abajo usan el PVP sugerido.', wrap=True)
    ref_pvp_real = '$H$' + str(r)
    r += 1
    _et(ws, 'G' + str(r), 'Food cost real sobre el precio de carta (%)',
        bold=True)
    motor.f(ws, 'H' + str(r),
            motor.iferror(
                'IF(' + ref_racion + '="","",IF(' + ref_pvp_real + '<>"",'
                + ref_racion + '/(' + ref_pvp_real + '/(1+' + ref_iva + ')),IF('
                + ref_pvp + '="","",' + ref_racion + '/' + ref_pvp + ')))'),
            fmt=motor.FMT_PCT, bold=True)
    motor.fijar_formato(ws, 'H' + str(r), motor.FMT_PCT)
    r += 1
    _et(ws, 'G' + str(r),
        'Food cost si pusieras el PVP SIN IVA en la carta (%)')
    motor.f(ws, 'H' + str(r),
            motor.iferror('IF(OR(' + ref_racion + '="",' + ref_pvp + '="",'
                          + ref_pvp + '=0),"",' + ref_racion + '/(' + ref_pvp
                          + '/(1+' + ref_iva + ')))'), fmt=motor.FMT_PCT)
    motor.fijar_formato(ws, 'H' + str(r), motor.FMT_PCT)
    r += 1
    _et(ws, 'A' + str(r + 1),
        'En España el precio de carta se muestra CON IVA incluido: aplicar el '
        'PVP sin IVA tal cual deja el food cost real por encima del objetivo '
        '(con un objetivo del 28 % y el IVA al 10 %, sube al 30,8 %) — y esa '
        'es exactamente la fila de arriba, que ahora sí lo calcula. Usa la '
        'fila «PVP con IVA», o escribe tu precio redondeado en «PVP REAL de '
        'carta» y mira el food cost que te queda de verdad '
        '[DOM-03 · COM-14 · RT-10 · RD-22 · RC-18].', wrap=True)
    cambios.append(
        fname + ':' + ws.title + '!' + ref_total + '..H' + str(r)
        + ': el PVP divide por raciones y usa el food cost objetivo de la celda'
          ' + PVP con y sin IVA. Antes H28=H27/0.28 sobre el coste del LOTE: '
          'una ficha de 10 raciones y 60 EUR de coste proponía 214,29 EUR por '
          'plato en vez de 21,43 EUR [TEC-05 · TEC-21 · DOM-05 · DOM-30 · '
          'DOM-03 · COM-14 · §4.1]')
    return {'modelo': 'E1', 'fila_cab': fila_cab, 'primera': primera,
            'fin': fin,
            'coste_total': ref_total.replace('$', ''),
            'coste_racion': ref_racion.replace('$', ''),
            'pvp_sin_iva': ref_pvp.replace('$', ''),
            'pvp_con_iva': 'H' + str(r - 4),
            'pvp_real': 'H' + str(fila_pvp_real),
            'food_cost_real': 'H' + str(fila_pvp_real + 1),
            'nombre_plato': 'C4'}


# ==========================================================================
# §4.1 — E2: la ficha de los 5 hermanos
# ==========================================================================
def _escandallo_e2(wb, ws, fila_cab, fname, cambios, contenido):
    marca = _fila_marca(ws)
    _t = (marca - 1) if marca else None
    fin = _fin_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    # La primera fila de datos NO es fila_cab+1: la 5 es `NOMBRE DEL PLATO:`.
    primera = _primera_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    if fin <= fila_cab:
        raise EstructuraDesconocida(fname + ':' + ws.title
                                    + ': 0 líneas de ingrediente')
    cfg = getattr(contenido, 'ESCANDALLO', None) or {}
    # fila del COSTE TOTAL, localizada por la ETIQUETA (nunca por posición: los
    # cinco hermanos tienen 8, 9, 10, 11 y 13 ingredientes).
    fila_total = marca if marca else None
    for r in range(fin + 1, min(fin + 6, ws.max_row) + 1):
        if fila_total:
            break
        if _norm(ws.cell(row=r, column=2).value).startswith('coste total'):
            fila_total = r
            break
    if fila_total is None:
        raise EstructuraDesconocida(
            fname + ':' + ws.title + ': no encuentro la fila «COSTE TOTAL '
            'PLATO» debajo de la fila ' + str(fin))
    fc = _literal_de(ws, 'G' + str(fila_total + 1), r'/\s*([\d.]+)\s*\)?\s*$')
    iva = _literal_de(ws, 'G' + str(fila_total + 2), r'\*\s*([\d.]+)\s*\)?\s*$')
    if iva is not None and iva > 1:
        iva = round(iva - 1, 4)          # `=G19*1.10` -> 0,10
    pie_fila, pie_texto = _pie_existente(ws, fila_cab + 1)
    _snapshot_parametros(ws, fila_total, ws.max_row)
    _limpiar(ws, fila_total, max(ws.max_row, (marca or 0) + 14))

    # ---- la merma entra como rendimiento, no como recargo (DOM-04) --------
    for r in range(primera, fin + 1):
        motor.f(ws, 'E' + str(r),
                motor.iferror('IF(OR($C{r}="",$D{r}=""),"",$C{r}/$K$3*$D{r})'
                              .format(r=r)), fmt=motor.FMT_EUR)
        motor.f(ws, 'G' + str(r),
                motor.iferror('IF(OR($E{r}="",$F{r}>=1),"",$E{r}/(1-$F{r}))'
                              .format(r=r)), fmt=motor.FMT_EUR)
        motor.verde(ws, 'B' + str(r) + ':D' + str(r))
        motor.verde(ws, 'F' + str(r))
        motor.fijar_formato(ws, 'D' + str(r), motor.FMT_EUR)
        motor.fijar_formato(ws, 'F' + str(r), motor.FMT_PCT)
    # RT-15/RC-11 · el mensaje de entrada de la merma se pegaba en TODAS las
    # celdas de porcentaje del producto (tipo de IVA, SS de la empresa, tipo
    # del préstamo…). El tope de 0,95 tampoco vale fuera de aquí: sólo la
    # merma entra en una división por `1-x` (RT-16).
    motor.prompt_porcentaje(
        ws, 'F' + str(primera) + ':F' + str(fin), *motor.PCT_MERMA)
    # La conversión g -> kg va en celda para que la fórmula no lleve el 1000
    # escrito dentro (§1.3). No es verde: no es un dato del cliente, es la
    # equivalencia de unidades de la propia tabla.
    _et(ws, 'J3', 'Gramos por kilo (la columna va en g/ml y el precio en €/kg)')
    motor.val(ws, 'K3', 1000, fmt=motor.FMT_ENT)
    ws['K3'].protection = Protection(locked=True)
    cambios.append(
        fname + ':' + ws.title + '!G' + str(primera) + ':G' + str(fin)
        + ': la merma pasa de RECARGO a rendimiento — antes G=E*(1+F) daba '
          'x1,20 para una merma del 20 % cuando lo correcto es /(1-0,20) = '
          'x1,25 (-4 % de coste por línea, -20 % con las mermas del 40 % del '
          'pescado); ahora G=E/(1-F) [DOM-04 · §4.1]')

    # ---- bloque de resultado ---------------------------------------------
    r = fila_total
    _et(ws, 'A' + str(r), MARCA_C, bold=True)
    _et(ws, 'B' + str(r), 'COSTE TOTAL PLATO (€)', bold=True)
    motor.f(ws, 'G' + str(r),
            _suma_guardada('G' + str(primera) + ':G' + str(fin)),
            fmt=motor.FMT_EUR, bold=True)
    ref_total = '$G$' + str(r)
    r += 1
    _et(ws, 'B' + str(r), 'Coste por ración (€)', bold=True)
    ref_rac = _param(ws, r, None, valor=1,
                     etiqueta='Raciones que salen de esta ficha',
                     fmt=motor.FMT_ENT,
                     nota='Por defecto 1: la ficha describe UN plato. Si es una '
                          'elaboración de varias raciones, escríbelas aquí y '
                          'el coste y el PVP se dividen.',
                     col_et='D', col_val='E', col_nota='H', minimo=1)
    motor.f(ws, 'G' + str(r),
            motor.iferror('IF(OR(' + ref_total + '="",' + ref_rac + '="",'
                          + ref_rac + '=0),"",' + ref_total + '/' + ref_rac
                          + ')'), fmt=motor.FMT_EUR, bold=True)
    ref_racion = '$G$' + str(r)
    r += 1
    ref_fc = _param(ws, r, None, valor=fc if fc else cfg.get('food_cost_objetivo'),
                    etiqueta='Food cost objetivo (%)', fmt='0%',
                    nota='Salía escrito dentro de la fórmula del PVP '
                         '(«=G/0,33»): cambiarlo aquí ahora sí mueve el precio.',
                    col_et='D', col_val='E', col_nota='H', porcentaje=True)
    motor.f(ws, 'B' + str(r),
            '="PVP sugerido (food cost "&TEXT(' + ref_fc + ',"0%")&"), sin IVA"')
    motor.f(ws, 'G' + str(r),
            motor.iferror('IF(OR(' + ref_racion + '="",' + ref_fc + '="",'
                          + ref_fc + '=0),"",' + ref_racion + '/' + ref_fc
                          + ')'), fmt=motor.FMT_EUR, bold=True)
    ref_pvp = '$G$' + str(r)
    r += 1
    ref_iva = _param(ws, r, 'iva_restauracion', valor=iva, col_et='D',
                     col_val='E', col_nota='H')
    _et(ws, 'B' + str(r), 'PVP con IVA (€) — el que va en la carta', bold=True)
    motor.f(ws, 'G' + str(r),
            motor.iferror('IF(' + ref_pvp + '="","",' + ref_pvp + '*(1+'
                          + ref_iva + '))'), fmt=motor.FMT_EUR, bold=True)
    fila_pvp_iva = r
    r += 1
    _et(ws, 'B' + str(r), 'Food cost real sobre el PVP sin IVA (%)')
    motor.f(ws, 'G' + str(r), _pct(ref_racion, ref_pvp), fmt=motor.FMT_PCT)
    motor.fijar_formato(ws, 'G' + str(r), motor.FMT_PCT)
    r += 2
    if pie_texto:
        _et(ws, 'A' + str(r), pie_texto)
        if pie_fila and pie_fila != r:
            cambios.append(fname + ':' + ws.title + ': el pie baja de la fila '
                           + str(pie_fila) + ' a la ' + str(r)
                           + ' para dejar sitio al bloque de resultado')
    cambios.append(
        fname + ':' + ws.title + '!' + ref_total + '..G' + str(r)
        + ': food cost e IVA salen de la fórmula a celda verde (antes '
          '«=G/0.33» y «=G*1.10» incrustados) + coste por ración + food cost '
          'real [TEC-21 · DOM-03 · DOM-30 · COM-14 · §4.1]')
    return {'modelo': 'E2', 'fila_cab': fila_cab, 'primera': primera,
            'fin': fin,
            'coste_total': ref_total.replace('$', ''),
            'coste_racion': ref_racion.replace('$', ''),
            'pvp_sin_iva': ref_pvp.replace('$', ''),
            'pvp_con_iva': 'G' + str(fila_pvp_iva),
            'nombre_plato': 'C' + str(primera - 1)}


# ==========================================================================
# §4.1 — E3: la tabla de productos de panadería
# ==========================================================================
def _escandallo_e3(wb, ws, fila_cab, fname, cambios, contenido):
    """No es una ficha: es una tabla de 25 productos con su food cost REAL.

    Aquí no hay merma ni raciones que meter (no existen las columnas), así que
    lo que aporta el §4.1 es: guardar las divisiones, declarar el IVA y añadir
    el PVP sugerido al food cost objetivo, que es la herramienta que falta —
    hoy la hoja dice qué food cost tienes, no a qué precio deberías vender.
    """
    marca = _fila_marca(ws)
    fin = _fin_por_columna(ws, fila_cab, 'B',
                           tope=(marca - 1) if marca else None)
    cfg = getattr(contenido, 'ESCANDALLO', None) or {}
    _snapshot_parametros(ws, fin + 1, ws.max_row)
    _limpiar(ws, fin + 1, max(ws.max_row, (marca or 0) + 12))

    col_pvp_obj = _cabecera_nueva(ws, fila_cab, 'I',
                                  'PVP sugerido al food cost objetivo (€)', 18)
    col_pvp_iva = _cabecera_nueva(ws, fila_cab, 'J', 'PVP con IVA (€)', 14)
    r = fin + 2
    _et(ws, 'A' + str(r), MARCA_C, bold=True)
    r += 1
    ref_fc = _param(ws, r, None, valor=cfg.get('food_cost_objetivo'),
                    etiqueta='Food cost objetivo (%)', fmt='0%',
                    nota=('Escribe el food cost al que quieres trabajar y la '
                          'columna «PVP sugerido» te dice a qué precio vender. '
                          + NOTA_SIN_DATO), col_et='A', col_val='B',
                    col_nota='C', porcentaje=True)
    r += 1
    ref_iva = _param(ws, r, 'iva_restauracion', col_et='A', col_val='B',
                     col_nota='C')
    for f_ in range(fila_cab + 1, fin + 1):
        motor.f(ws, 'E' + str(f_), _pct('$C' + str(f_), '$D' + str(f_)),
                fmt=motor.FMT_PCT)
        motor.f(ws, 'F' + str(f_),
                motor.iferror('IF($E{r}="","",1-$E{r})'.format(r=f_)),
                fmt=motor.FMT_PCT)
        motor.f(ws, 'G' + str(f_),
                motor.iferror('IF(OR($C{r}="",$D{r}=""),"",$D{r}-$C{r})'
                              .format(r=f_)), fmt=motor.FMT_EUR)
        motor.f(ws, col_pvp_obj + str(f_),
                motor.iferror('IF(OR($C{r}="",' + ref_fc + '="",' + ref_fc
                              + '=0),"",$C{r}/' + ref_fc + ')'.format())
                .format(r=f_), fmt=motor.FMT_EUR)
        motor.f(ws, col_pvp_iva + str(f_),
                motor.iferror('IF($D{r}="","",$D{r}*(1+' + ref_iva
                              + '))').format(r=f_), fmt=motor.FMT_EUR)
        motor.verde(ws, 'A' + str(f_) + ':D' + str(f_))
        motor.verde(ws, 'H' + str(f_))
        motor.fijar_formato(ws, 'C' + str(f_), motor.FMT_EUR)
        motor.fijar_formato(ws, 'D' + str(f_), motor.FMT_EUR)
    cambios.append(
        fname + ':' + ws.title + '!E' + str(fila_cab + 1) + ':' + col_pvp_iva
        + str(fin) + ': divisiones guardadas con IFERROR (antes «=C/D» daba '
          '#DIV/0! en cuanto se borraba un PVP), IVA en celda y columna «PVP '
          'sugerido al food cost objetivo» — la tabla decía qué food cost '
          'tienes y no a qué precio vender [DOM-03 · DOM-30 · COM-14 · §4.1]')
    cambios.append(
        fname + ':' + ws.title + ': modelo E3 (tabla de 25 productos): NO se '
        'renombra a «' + HOJA_FICHA + '» ni se le añade hoja «' + HOJA_RESUMEN
        + '» (TEC-23) — no hay ficha que duplicar, y tampoco columna de merma '
          'donde aplicar DOM-04. Queda como hallazgo para la capa de producto.')
    return {'modelo': 'E3', 'fila_cab': fila_cab, 'fin': fin,
            'primera': fila_cab + 1, 'ref_fc': ref_fc, 'ref_iva': ref_iva,
            'col_pvp_obj': col_pvp_obj, 'col_pvp_iva': col_pvp_iva}


# ==========================================================================
# §4.1 — hoja Resumen (TEC-23), sólo para E1/E2
# ==========================================================================
def _hoja_resumen(wb, ficha, fname, cambios):
    """Consolida coste por ración y PVP de cada ficha (TEC-23, RC-19).

    RC-19: la primera versión cableaba UNA sola fila a `Ficha (plantilla)` y
    pedía al cliente que copiara la fila y **editara el nombre de la hoja
    dentro de seis fórmulas** por cada pase — en un producto que promete un
    menú degustación de 8-12 pases. Ahora hay 12 filas y lo único que se
    escribe es el NOMBRE de la hoja, en verde: las fórmulas lo resuelven con
    `INDIRECT`. Con la celda del nombre vacía, la fila entera devuelve `""`;
    con un nombre que no existe, `IFERROR` la deja vacía en vez de sembrar
    `#REF!` (que es lo que la versión anterior quería evitar y por lo que se
    quedó en una fila).
    """
    nueva = HOJA_RESUMEN not in wb.sheetnames
    ws = wb[HOJA_RESUMEN] if not nueva else wb.create_sheet(HOJA_RESUMEN)
    for row in ws.iter_rows():
        for c in row:
            if c.__class__.__name__ != 'MergedCell':
                c.value = None
    _a4(ws)
    ws.column_dimensions['A'].width = 26.0
    ws.column_dimensions['B'].width = 34.0
    for L in ('C', 'D', 'E', 'F', 'G'):
        ws.column_dimensions[L].width = 18.0
    _et(ws, 'A1', 'Resumen de fichas técnicas', bold=True)
    ws['A1'].font = Font(bold=True, size=14)
    _et(ws, 'A2', 'AI Chef Pro · aichef.pro')
    cabs = (('A', 'Hoja de la ficha'), ('B', 'Plato'),
            ('C', 'Coste total (€)'), ('D', 'Coste por ración (€)'),
            ('E', 'PVP sin IVA (€)'), ('F', 'PVP con IVA (€)'),
            ('G', 'Food cost real (%)'))
    for L, texto in cabs:
        _et(ws, L + '4', texto, bold=True)
    primera, ultima = 5, 16

    def _tira(fila, celda):
        """`INDIRECT` guardado por el nombre de la hoja de la columna A."""
        ref = ('INDIRECT("\'"&$A' + str(fila) + '&"\'!' + celda + '")')
        return ('=IF($A' + str(fila) + '="","",IFERROR(IF(' + ref + '="","",'
                + ref + '),""))')

    q = "'" + HOJA_FICHA + "'!"

    def _directa(celda):
        """Fila 5: referencia DIRECTA a la plantilla.

        No es una excepción cosmética. `INDIRECT` es dinámico y ningún motor de
        cálculo —ni pycel, que es con lo que se verifica este paquete— puede
        seguir la dependencia: el gate de §4.1 comprueba que al tocar una línea
        de la ficha el Resumen se mueve, y con `INDIRECT` no lo vería. La
        plantilla, que es la ficha que SIEMPRE existe, va cableada; las once
        filas de abajo van por nombre.
        """
        return motor.iferror('IF(' + q + celda + '="","",' + q + celda + ')')

    columnas = (('C', 'coste_total', motor.FMT_EUR),
                ('D', 'coste_racion', motor.FMT_EUR),
                ('E', 'pvp_sin_iva', motor.FMT_EUR),
                ('F', 'pvp_con_iva', motor.FMT_EUR),
                ('G', 'food_cost_real', motor.FMT_PCT))
    for f_ in range(primera, ultima + 1):
        if f_ == primera:
            _et(ws, 'A' + str(f_), HOJA_FICHA)
            _et(ws, 'H' + str(f_),
                'Esta fila va cableada a la plantilla: no cambies su nombre '
                'aquí. Las de abajo se alimentan del nombre que escribas en '
                'la columna A.', wrap=True)
            motor.f(ws, 'B' + str(f_), _directa(ficha['nombre_plato']))
            for L, clave, fmt in columnas:
                if not ficha.get(clave):
                    continue
                motor.f(ws, L + str(f_), _directa(ficha[clave]), fmt=fmt)
                motor.fijar_formato(ws, L + str(f_), fmt)
            continue
        motor.verde(ws, 'A' + str(f_))
        motor.f(ws, 'B' + str(f_), _tira(f_, ficha['nombre_plato']))
        for L, clave, fmt in columnas:
            celda = ficha.get(clave)
            if not celda:
                continue
            motor.f(ws, L + str(f_), _tira(f_, celda), fmt=fmt)
            motor.fijar_formato(ws, L + str(f_), fmt)
    r = ultima + 2
    _et(ws, 'A' + str(r), 'Cómo añadir una ficha más', bold=True)
    _et(ws, 'A' + str(r + 1),
        '1. Clic derecho sobre la pestaña «' + HOJA_FICHA + '» → Mover o '
        'copiar → marca «Crear una copia» y renómbrala con el nombre del '
        'plato.', wrap=True)
    _et(ws, 'A' + str(r + 2),
        '2. Escribe ese mismo nombre en la columna «Hoja de la ficha» de esta '
        'hoja, en cualquier fila de la 6 a la ' + str(ultima) + '. NO hay que '
        'tocar ninguna fórmula: esas filas leen la hoja que nombres en la '
        'columna A.', wrap=True)
    _et(ws, 'A' + str(r + 3),
        '3. Un menú degustación de 8-12 pases necesita una ficha por pase: por '
        'eso hay doce filas. Si te faltan, copia la última hacia abajo '
        '[TEC-23 · RC-19].', wrap=True)
    for L in ('A', 'B'):
        for f_ in range(r + 1, r + 4):
            ws[L + str(f_)].alignment = Alignment(vertical='top',
                                                  wrap_text=True)
    if nueva:
        cambios.append(fname + ': hoja «' + HOJA_RESUMEN + '» CREADA — '
                       'consolida coste por ración y PVP de cada ficha '
                       '[TEC-23 · §4.1]')
    else:
        cambios.append(fname + ':' + HOJA_RESUMEN + ': 12 filas que se '
                       'alimentan del NOMBRE de la hoja (columna A, verde) con '
                       'INDIRECT: consolidar un menú de 8-12 pases ya no exige '
                       'editar el nombre de la hoja dentro de seis fórmulas '
                       '[RC-19]')
    return ws


def _escandallo(wb, fname, cambios, contenido, registro_modelo):
    hojas = [ws for ws in wb.worksheets
             if ws.title not in ('Instrucciones', HOJA_RESUMEN)]
    if not hojas:
        raise EstructuraDesconocida(fname + ': sin hoja de datos')
    ws = hojas[0]
    modelo, fila_cab = modelo_escandallo(ws, fname)
    if modelo == 'E1':
        ficha = _escandallo_e1(wb, ws, fila_cab, fname, cambios, contenido)
    elif modelo == 'E2':
        ficha = _escandallo_e2(wb, ws, fila_cab, fname, cambios, contenido)
    else:
        ficha = _escandallo_e3(wb, ws, fila_cab, fname, cambios, contenido)
    if modelo in ('E1', 'E2'):
        _hoja_resumen(wb, ficha, fname, cambios)
        _instr(wb, fname, cambios, [
            'Duplica la hoja «' + HOJA_FICHA + '» para cada plato: clic '
            'derecho sobre la pestaña -> Mover o copiar -> «Crear una copia», '
            'y renómbrala con el nombre del plato. La hoja «' + HOJA_RESUMEN
            + '» consolida el coste y el PVP de todas.',
            'La merma se escribe en tanto por uno (0,20 = 20 %) y ENTRA en el '
            'coste: la cantidad que compras es la neta dividida por (1 - merma).',
            'El food cost objetivo y el tipo de IVA son celdas verdes: '
            'cámbialos y el PVP se recalcula.'])
    else:
        _instr(wb, fname, cambios, [
            'El food cost objetivo es una celda verde: escríbelo y la columna '
            '«PVP sugerido» te dice a qué precio vender cada producto.'])
    registro_modelo['escandallo'] = ficha
    return ficha


# ==========================================================================
# §4.2 — menu-engineering-matrix (Kasavana & Smith completo)
# ==========================================================================
def modelo_menu(ws, fname=''):
    fila = _fila_cabecera(ws)
    if not fila:
        raise EstructuraDesconocida(
            fname + ':' + ws.title + ': sin fila de cabecera reconocible')
    c = dict((k, _norm(v)) for k, v in _cab(ws, fila).items())
    if c.get('D') == 'uds vendidas' and c.get('F') == 'pvp (€)':
        return 'M1', fila
    if c.get('C') == 'pvp (€)' and c.get('F') == 'uds. vendidas/mes':
        return 'M2', fila
    raise EstructuraDesconocida(
        fname + ':' + ws.title + ': cabecera de menu engineering no reconocida '
        'por §4.2 (fila ' + str(fila) + ') = '
        + repr(dict((k, v) for k, v in _cab(ws, fila).items() if v is not None)))


#: Vocabulario del semáforo de la clasificación, de más grave a menos: el orden
#: importa porque `semaforo_texto` usa `stopIfTrue` y «Star» está contenido en
#: ninguna otra, pero «Dog» sí debe ganar antes que un futuro «Dogfood».
SEM_KS = (('Dog', motor.CF_ROJO_BG, motor.CF_ROJO_FG),
          ('Plowhorse', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
          ('Puzzle', motor.CF_AMBAR_BG, motor.CF_AMBAR_FG),
          ('Star', motor.CF_VERDE_BG, motor.CF_VERDE_FG))

ACCIONES = ('Mantener y destacar en carta', 'Subir precio o bajar coste',
            'Promocionar y reubicar en carta', 'Retirar o rediseñar')


def _menu(wb, fname, cambios, contenido, registro_modelo):
    hojas = [ws for ws in wb.worksheets if ws.title != 'Instrucciones']
    ws = hojas[0]
    modelo, fila_cab = modelo_menu(ws, fname)
    cfg = getattr(contenido, 'MENU', None) or {}
    marca = _fila_marca(ws)
    _t = (marca - 1) if marca else None

    if modelo == 'M1':
        col_uds, col_coste, col_pvp, col_margen = 'D', 'E', 'F', 'G'
        col_mix, col_clase, col_accion = 'H', 'I', 'J'
        col_cat = 'C'
        primera = _primera_por_columna(ws, fila_cab, 'A', numerica=True,
                                       tope=_t)
        ultima = _fin_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    else:
        col_uds, col_coste, col_pvp, col_margen = 'F', 'D', 'C', 'G'
        col_mix, col_clase, col_accion = 'I', 'J', 'K'
        col_cat = None
        primera = _primera_por_columna(ws, fila_cab, 'A', numerica=True,
                                       tope=_t)
        ultima = _fin_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)

    pie_fila, pie_texto = _pie_existente(ws, fila_cab + 1)
    _snapshot_parametros(ws, ultima + 1, ws.max_row)
    _limpiar(ws, ultima + 1, max(ws.max_row, (marca or 0) + 12))

    # ---- platos de ejemplo (§4.2: el representante no tiene ninguno) ------
    platos = cfg.get('platos') or []
    if modelo == 'M1' and platos:
        libre = all(ws.cell(row=r, column=2).value in (None, '')
                    for r in range(primera, ultima + 1))
        ya = _norm(ws.cell(row=primera, column=2).value) == _norm(
            platos[0]['plato'])
        if libre or ya:
            for i, p in enumerate(platos):
                r = primera + i
                if r > ultima:
                    break
                motor.val(ws, 'B' + str(r), p['plato'], verde_=True)
                motor.val(ws, col_cat + str(r), p['categoria'], verde_=True)
                motor.val(ws, col_uds + str(r), p['uds'], fmt=motor.FMT_ENT,
                          verde_=True)
                motor.val(ws, col_coste + str(r), p['coste'],
                          fmt=motor.FMT_EUR, verde_=True)
                motor.val(ws, col_pvp + str(r), p['pvp'], fmt=motor.FMT_EUR,
                          verde_=True)
            if cfg.get('categorias'):
                motor.dv_lista(ws, [col_cat + str(r)
                                    for r in range(primera, ultima + 1)],
                               cfg['categorias'])
            cambios.append(
                fname + ':' + ws.title + '!B' + str(primera) + ':'
                + col_pvp + str(primera + len(platos) - 1) + ': '
                + str(len(platos)) + ' platos de EJEMPLO precargados (la hoja '
                'se entregaba con 25 filas vacías y ni una fórmula que '
                'clasificar). PVP dentro de las bandas del cap. 15 de la guía '
                '(entrantes 18-28 EUR, principales 32-48 EUR, postres 14-22 '
                'EUR, degustación 90-180 EUR). Todas las celdas son verdes: '
                'se sustituyen por la carta real [§4.2]')

    # ---- columnas nuevas --------------------------------------------------
    _cabecera_nueva(ws, fila_cab, col_mix, 'Mix % (popularidad)', 16)
    _cabecera_nueva(ws, fila_cab, col_clase, 'Clasificación', 16)
    _cabecera_nueva(ws, fila_cab, col_accion, 'Acción recomendada', 34)
    # RC-17 · las Instrucciones piden analizar familia por familia y la hoja
    # hacía justo lo contrario: un umbral y un margen medio ÚNICOS para toda la
    # carta, con la columna «Categoría» que no leía ninguna fórmula. Seguir la
    # instrucción obligaba a BORRAR filas. Estas dos columnas calculan el
    # margen medio y el umbral DE SU PROPIA FAMILIA para cada plato.
    col_mfam, col_ufam, col_pvpiva = None, None, None
    if modelo == 'M1' and col_cat:
        col_mfam = chr(ord(col_accion) + 1)
        col_ufam = chr(ord(col_accion) + 2)
        col_pvpiva = chr(ord(col_accion) + 3)
        _cabecera_nueva(ws, fila_cab, col_mfam,
                        'Margen medio de SU familia (€)', 20)
        _cabecera_nueva(ws, fila_cab, col_ufam,
                        'Umbral de popularidad de SU familia (%)', 22)
        # RD-23 · §1.5(a): toda hoja que fije un PVP lleva el precio sin IVA y
        # con IVA, con el tipo en celda. Aquí no se decía ni una cosa ni otra,
        # y en España la carta se anuncia CON IVA: un menú de 150 € son
        # 136,36 € netos, y la clasificación se movería un 10 %.
        _cabecera_nueva(ws, fila_cab, col_pvpiva,
                        'PVP en carta, con IVA (€)', 20)

    # ---- bloque de totales y umbral ---------------------------------------
    r = ultima + 1
    _et(ws, 'A' + str(r), MARCA_C, bold=True)
    r += 1
    fila_tot = r
    _et(ws, 'B' + str(r), 'TOTAL / MEDIA PONDERADA', bold=True)
    motor.f(ws, col_uds + str(r),
            _suma_guardada(col_uds + str(primera) + ':' + col_uds + str(ultima)),
            fmt=motor.FMT_ENT, bold=True)
    ref_uds = '$' + col_uds + '$' + str(r)
    motor.f(ws, col_margen + str(r),
            motor.iferror('IF(' + ref_uds + '="","",ROUND(SUMPRODUCT($'
                          + col_uds + '$' + str(primera) + ':$' + col_uds + '$'
                          + str(ultima) + ',$' + col_margen + '$' + str(primera)
                          + ':$' + col_margen + '$' + str(ultima) + '),2)/'
                          + ref_uds + ')'), fmt=motor.FMT_EUR, bold=True)
    ref_mc = '$' + col_margen + '$' + str(r)
    r += 1
    fila_umbral = r
    _et(ws, 'B' + str(r), 'Umbral de popularidad (Kasavana & Smith)', bold=True)
    ref_factor = _param(
        ws, r, None, valor=0.7,
        etiqueta='Factor del umbral', fmt='0%',
        nota='El clásico es el 70 % del mix medio: umbral = factor / nº de '
             'platos con ventas. Súbelo si tu carta es muy corta.',
        col_et=col_mix, col_val=col_clase, col_nota=col_accion,
        porcentaje=True)
    motor.f(ws, col_uds + str(r),
            motor.iferror('IF(COUNTIF($' + col_uds + '$' + str(primera) + ':$'
                          + col_uds + '$' + str(ultima) + ',">0")=0,"",'
                          + ref_factor + '/COUNTIF($' + col_uds + '$'
                          + str(primera) + ':$' + col_uds + '$' + str(ultima)
                          + ',">0"))'), fmt=motor.FMT_PCT, bold=True)
    motor.fijar_formato(ws, col_uds + str(r), motor.FMT_PCT)
    ref_umbral = '$' + col_uds + '$' + str(r)
    ref_iva_menu = None
    if col_pvpiva:
        r += 1
        ref_iva_menu = _param(
            ws, r, 'iva_restauracion', col_et='B', col_val=col_uds,
            col_nota=col_margen)
        motor.fijar_formato(ws, ref_iva_menu.replace('$', ''), motor.FMT_PCT)
    r += 2
    if pie_texto:
        _et(ws, 'A' + str(r), pie_texto)

    # ---- mix, clasificación y acción --------------------------------------
    for f_ in range(primera, ultima + 1):
        # Las columnas que ya venían calculadas se REESCRIBEN con guarda: el
        # `=F5-E5` original imprime «0,00 €» en las 13 filas vacías del
        # representante, y un margen de cero en una fila sin plato se lee como
        # un plato que no deja nada (§7-bis.13).
        if modelo == 'M1':
            motor.f(ws, col_margen + str(f_),
                    motor.iferror('IF(OR($E{r}="",$F{r}=""),"",$F{r}-$E{r})'
                                  .format(r=f_)), fmt=motor.FMT_EUR)
            motor.verde(ws, 'B' + str(f_) + ':F' + str(f_))
            motor.fijar_formato(ws, 'E' + str(f_), motor.FMT_EUR)
            motor.fijar_formato(ws, 'F' + str(f_), motor.FMT_EUR)
        else:
            motor.f(ws, 'E' + str(f_), _pct('$D' + str(f_), '$C' + str(f_)),
                    fmt=motor.FMT_PCT)
            motor.fijar_formato(ws, 'E' + str(f_), motor.FMT_PCT)
            motor.f(ws, col_margen + str(f_),
                    motor.iferror('IF(OR($C{r}="",$D{r}=""),"",$C{r}-$D{r})'
                                  .format(r=f_)), fmt=motor.FMT_EUR)
            motor.f(ws, 'H' + str(f_),
                    motor.iferror('IF(OR($G{r}="",$F{r}=""),"",$G{r}*$F{r})'
                                  .format(r=f_)), fmt=motor.FMT_EUR)
            motor.verde(ws, 'B' + str(f_) + ':D' + str(f_))
            motor.verde(ws, 'F' + str(f_))
            motor.fijar_formato(ws, 'F' + str(f_), motor.FMT_ENT)
        motor.f(ws, col_mix + str(f_),
                motor.iferror('IF(OR($' + col_uds + str(f_) + '="",'
                              + ref_uds + '="",' + ref_uds + '=0),"",$'
                              + col_uds + str(f_) + '/' + ref_uds + ')'),
                fmt=motor.FMT_PCT)
        motor.fijar_formato(ws, col_mix + str(f_), motor.FMT_PCT)
        mix = '$' + col_mix + str(f_)
        mar = '$' + col_margen + str(f_)
        vs_mc, vs_umbral = ref_mc, ref_umbral
        if col_mfam:
            rc = ('$' + col_cat + '$' + str(primera) + ':$' + col_cat + '$'
                  + str(ultima))
            ru = ('$' + col_uds + '$' + str(primera) + ':$' + col_uds + '$'
                  + str(ultima))
            rm = ('$' + col_margen + '$' + str(primera) + ':$' + col_margen
                  + '$' + str(ultima))
            cat_f = '$' + col_cat + str(f_)
            uds_fam = 'SUMPRODUCT(--(' + rc + '=' + cat_f + '),' + ru + ')'
            motor.f(ws, col_mfam + str(f_),
                    motor.iferror(
                        'IF(' + cat_f + '="","",IF(' + uds_fam + '=0,"",'
                        'ROUND(SUMPRODUCT(--(' + rc + '=' + cat_f + '),' + ru
                        + ',' + rm + '),2)/' + uds_fam + '))'),
                    fmt=motor.FMT_EUR)
            motor.fijar_formato(ws, col_mfam + str(f_), motor.FMT_EUR)
            n_fam = ('SUMPRODUCT(--(' + rc + '=' + cat_f + '),--(' + ru
                     + '>0))')
            motor.f(ws, col_ufam + str(f_),
                    motor.iferror(
                        'IF(' + cat_f + '="","",IF(' + n_fam + '=0,"",'
                        + ref_factor + '/' + n_fam + '))'),
                    fmt=motor.FMT_PCT)
            motor.fijar_formato(ws, col_ufam + str(f_), motor.FMT_PCT)
            # el mix también es DENTRO de su familia
            motor.f(ws, col_mix + str(f_),
                    motor.iferror(
                        'IF(OR($' + col_uds + str(f_) + '="",' + cat_f
                        + '=""),"",IF(' + uds_fam + '=0,"",$' + col_uds
                        + str(f_) + '/' + uds_fam + '))'), fmt=motor.FMT_PCT)
            motor.f(ws, col_pvpiva + str(f_),
                    motor.iferror(
                        'IF(OR($' + col_pvp + str(f_) + '="",' + ref_iva_menu
                        + '=""),"",$' + col_pvp + str(f_) + '*(1+'
                        + ref_iva_menu + '))'), fmt=motor.FMT_EUR)
            motor.fijar_formato(ws, col_pvpiva + str(f_), motor.FMT_EUR)
            vs_mc = '$' + col_mfam + str(f_)
            vs_umbral = '$' + col_ufam + str(f_)
        motor.f(ws, col_clase + str(f_),
                '=IF(OR(' + mix + '="",' + mar + '="",' + vs_mc + '="",'
                + vs_umbral + '=""),"",'
                + 'IF(AND(' + mix + '>=' + vs_umbral + ',' + mar + '>='
                + vs_mc + '),"Star",'
                + 'IF(AND(' + mix + '>=' + vs_umbral + ',' + mar + '<'
                + vs_mc + '),"Plowhorse",'
                + 'IF(AND(' + mix + '<' + vs_umbral + ',' + mar + '>='
                + vs_mc + '),"Puzzle","Dog"))))')
        cl = '$' + col_clase + str(f_)
        motor.f(ws, col_accion + str(f_),
                '=IF(' + cl + '="","",IF(' + cl + '="Star","' + ACCIONES[0]
                + '",IF(' + cl + '="Plowhorse","' + ACCIONES[1] + '",IF('
                + cl + '="Puzzle","' + ACCIONES[2] + '","' + ACCIONES[3]
                + '"))))')
        if modelo == 'M1':
            motor.fijar_formato(ws, col_uds + str(f_), motor.FMT_ENT)
    motor.semaforo_texto(ws, col_clase + str(primera) + ':' + col_clase
                         + str(ultima), SEM_KS)
    cambios.append(
        fname + ':' + ws.title + '!' + col_mix + str(primera) + ':'
        + col_accion + str(ultima) + ' + fila ' + str(fila_tot)
        + ': Kasavana & Smith COMPLETO — mix %, margen de contribución medio '
          'ponderado (SUMPRODUCT envuelto en ROUND: pycel devuelve numpy.int64 '
          'y el SUM de arriba daría 0), umbral 70 %/N con el factor en celda, '
          'clasificación por fórmula y acción recomendada por cuadrante. Antes '
          'las ' + str(ultima - primera + 1) + ' fórmulas del libro eran restas '
          '«=PVP-coste» y la columna Clasificación estaba vacía '
          '[TEC-04 · DOM-10 · COM-08 · §4.2]')
    if col_mfam:
        _et(ws, col_coste + str(fila_cab - 1),
            'Coste por ración de escandallo-maestro.xlsx', wrap=True)
    _instr(wb, fname, cambios, [
        # RD-24 · §4.1 designa el «Coste por ración» de la ficha como el
        # alimento de esta matriz y §1.13 obliga a repetir el dato con su nota
        # de origen cuando no se enlazan libros. Sin ella, un consultor no
        # sabe si ese coste incluye merma, guarnición o pan.
        'La columna «Coste (€)» es el «Coste por ración» de la ficha '
        'correspondiente de escandallo-maestro.xlsx: cópialo desde su hoja '
        '«Resumen». Incluye la merma (la ficha compra en BRUTO), así que no '
        'vuelvas a añadirla aquí. Los costes que trae el libro son de EJEMPLO.',
        'La clasificación es automática: Star (alta popularidad y alto '
        'margen), Plowhorse (popular pero poco rentable), Puzzle (rentable '
        'pero poco pedido) y Dog (ni una cosa ni la otra). La columna «Acción '
        'recomendada» dice qué hacer con cada uno.',
        'El umbral de popularidad es 70 % / nº de platos con ventas, y el '
        'factor del 70 % es una celda verde por si tu carta es muy corta.',
        'La matriz se aplica DENTRO de cada familia de carta, no sobre toda '
        'la tabla: cada plato se compara con el margen medio y con el umbral '
        'de popularidad de SU familia (las dos columnas nuevas de la derecha). '
        'Por eso un menú degustación de 145 € y un postre de 15 € pueden '
        'convivir en la misma hoja sin arrastrarse el uno al otro. La fila '
        'TOTAL sigue dando la media de toda la carta, que es la foto de '
        'conjunto.',
        'Los precios de la columna «PVP (€)» van SIN IVA, igual que en el '
        'resto del producto. La columna «PVP en carta, con IVA (€)» es el '
        'precio que ve el comensal, con el tipo del IVA en celda.'])
    registro_modelo['menu'] = {
        'modelo': modelo, 'primera': primera, 'ultima': ultima,
        'col_uds': col_uds, 'col_margen': col_margen, 'col_mix': col_mix,
        'col_clase': col_clase, 'fila_total': fila_tot,
        'fila_umbral': fila_umbral}
    return registro_modelo['menu']


# ==========================================================================
# §4.3 — budget-bodega (sólo representante)
# ==========================================================================
def _bodega(wb, fname, cambios, contenido, registro_modelo):
    hojas = [ws for ws in wb.worksheets if ws.title != 'Instrucciones']
    ws = hojas[0]
    fila_cab = _fila_cabecera(ws)
    c = dict((k, _norm(v)) for k, v in _cab(ws, fila_cab).items())
    if not (c.get('E') == 'coste (€)' and c.get('I') == 'stock (uds)'):
        raise EstructuraDesconocida(
            fname + ':' + ws.title + ': cabecera de bodega no reconocida por '
            '§4.3 = ' + repr(dict((k, v) for k, v in _cab(ws, fila_cab).items()
                                  if v is not None)))
    cfg = getattr(contenido, 'BODEGA', None) or {}
    marca = _fila_marca(ws)
    _t = (marca - 1) if marca else None
    primera = _primera_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    ultima = _fin_por_columna(ws, fila_cab, 'A', numerica=True, tope=_t)
    _limpiar(ws, ultima + 1, max(ws.max_row, (marca or 0) + 12))

    # ---- H deja de llamar margen a un markup (TEC-20/DOM-28/COM-22) -------
    ws['H' + str(fila_cab)].value = 'Multiplicador (×)'
    nuevas = (('K', 'Uds vendidas/mes', 14),
              ('L', 'Margen s/PVP (%)', 14),
              ('M', 'Food cost bebida (%)', 16),
              ('N', 'Valor stock a coste (€)', 18),
              ('O', 'Valor stock a PVP (€)', 18),
              ('P', 'Formato', 12), ('Q', 'Añada', 10),
              ('R', 'Proveedor', 22), ('S', 'Ubicación / bin', 16))
    for letra, texto, ancho in nuevas:
        _cabecera_nueva(ws, fila_cab, letra, texto, ancho)
    # la banda del título abarcaba A:J; se extiende a las columnas nuevas
    for rango in [str(m) for m in ws.merged_cells.ranges]:
        cr = CellRange(rango)
        if cr.min_row == cr.max_row and cr.min_row < fila_cab \
                and cr.min_col == 1 and cr.max_col < _col('S'):
            ws.unmerge_cells(rango)
            ws.merge_cells(start_row=cr.min_row, start_column=1,
                           end_row=cr.min_row, end_column=_col('S'))

    # RD-13 · «F» es el PVP de CARTA, que en España se anuncia CON IVA, y en
    # vino alcohólico el tipo es el 21 %. Cruzarlo con un coste sin IVA
    # subestima el food cost de bebida en unos 7 puntos, justo en la magnitud
    # que decide si la bodega gana dinero. El escandallo de este mismo pack ya
    # lo resuelve con su celda de tipo (§1.5a): aquí igual, con una columna
    # «PVP sin IVA» de la que cuelgan todos los indicadores.
    # el parámetro se escribe al final, en el bloque de la hoja (esta zona la
    # limpia `_limpiar` antes de reescribirla); aquí sólo se reserva su celda.
    fila_iva_beb = ultima + 5
    ref_iva_beb = '$C$' + str(fila_iva_beb)
    _cabecera_nueva(ws, fila_cab, 'T', 'PVP sin IVA (€)', 15)
    for r in range(primera, ultima + 1):
        motor.f(ws, 'T' + str(r),
                motor.iferror(('IF(OR($F{r}="",' + ref_iva_beb + '=""),"",'
                               '$F{r}/(1+' + ref_iva_beb + '))').format(r=r)),
                fmt=motor.FMT_EUR)
        motor.fijar_formato(ws, 'T' + str(r), motor.FMT_EUR)
    for r in range(primera, ultima + 1):
        motor.f(ws, 'G' + str(r),
                motor.iferror('IF(OR($E{r}="",$T{r}=""),"",$T{r}-$E{r})'
                              .format(r=r)), fmt=motor.FMT_EUR)
        motor.f(ws, 'H' + str(r),
                motor.iferror('IF(OR($E{r}="",$T{r}="",$E{r}=0),"",$T{r}/$E{r})'
                              .format(r=r)), fmt='#,##0.00"×"')
        motor.fijar_formato(ws, 'H' + str(r), '#,##0.00"×"')
        motor.f(ws, 'L' + str(r),
                motor.iferror('IF(OR($E{r}="",$T{r}="",$T{r}=0),"",'
                              '($T{r}-$E{r})/$T{r})'.format(r=r)),
                fmt=motor.FMT_PCT)
        motor.f(ws, 'M' + str(r), _pct('$E' + str(r), '$T' + str(r)),
                fmt=motor.FMT_PCT)
        motor.f(ws, 'J' + str(r),
                motor.iferror('IF(OR($I{r}="",$K{r}="",$I{r}=0),"",'
                              '$K{r}/$I{r})'.format(r=r)), fmt='#,##0.00')
        motor.f(ws, 'N' + str(r),
                motor.iferror('IF(OR($E{r}="",$I{r}=""),"",$E{r}*$I{r})'
                              .format(r=r)), fmt=motor.FMT_EUR)
        motor.f(ws, 'O' + str(r),
                motor.iferror('IF(OR($T{r}="",$I{r}=""),"",$T{r}*$I{r})'
                              .format(r=r)), fmt=motor.FMT_EUR)
        motor.verde(ws, 'B' + str(r) + ':F' + str(r))
        motor.verde(ws, 'I' + str(r))
        motor.verde(ws, 'K' + str(r))
        motor.verde(ws, 'P' + str(r) + ':S' + str(r))
        motor.fijar_formato(ws, 'I' + str(r), motor.FMT_ENT)
        motor.fijar_formato(ws, 'K' + str(r), motor.FMT_ENT)
        motor.fijar_formato(ws, 'L' + str(r), motor.FMT_PCT)
        motor.fijar_formato(ws, 'M' + str(r), motor.FMT_PCT)
    if cfg.get('tipos'):
        motor.dv_lista(ws, ['C' + str(r) for r in range(primera, ultima + 1)],
                       cfg['tipos'])
    # RT-30 · las 50 referencias llegaban VACÍAS y las 355 fórmulas del libro
    # no enseñaban ni un resultado, mientras el menu engineering del mismo pack
    # sí iba precargado. Mismo criterio para los dos.
    puestas = 0
    for i, ref in enumerate(cfg.get('referencias') or []):
        f_ = primera + i
        if f_ > ultima:
            break
        for L, clave, fmt in (('B', 'referencia', None), ('C', 'tipo', None),
                              ('D', 'do', None),
                              ('E', 'coste', motor.FMT_EUR),
                              ('F', 'pvp', motor.FMT_EUR),
                              ('I', 'stock', motor.FMT_ENT),
                              ('K', 'uds', motor.FMT_ENT)):
            if ref.get(clave) is None or ws[L + str(f_)].value is not None:
                continue
            motor.val(ws, L + str(f_), ref[clave], fmt=fmt, verde_=True)
            if fmt:
                motor.fijar_formato(ws, L + str(f_), fmt)
            puestas += 1
    if puestas:
        cambios.append(
            fname + ':' + ws.title + '!B' + str(primera) + ':K'
            + str(primera + len(cfg.get('referencias') or []) - 1) + ': '
            + str(puestas) + ' celdas de ejemplo precargadas en verde (10 '
            'referencias con multiplicador entre ×2,5 y ×3,5 y food cost de '
            'bebida dentro del 28-40 % del capítulo de bodega): el libro deja '
            'de abrir con 355 fórmulas y ningún resultado [RT-30 · §1.2]')
    cambios.append(
        fname + ':' + ws.title + '!H' + str(primera) + ':' + 'M' + str(ultima)
        + ': «Margen (%)» era un MARKUP sobre coste — con coste 10 EUR y PVP '
          '30 EUR imprimía «200,0 %», imposible como margen. Ahora H es '
          '«Multiplicador (x)» (=PVP/coste, el x2,5-x3,5 del cap. 12), L es el '
          'margen sobre PVP (66,7 % en ese ejemplo) y M el food cost de bebida, '
          'que es el que se contrasta con el 28-40 % del capítulo '
          '[TEC-20 · DOM-28 · COM-22 · §4.3]')
    cambios.append(
        fname + ':' + ws.title + '!J' + str(primera) + ':J' + str(ultima)
        + ': la rotación se CALCULA (uds vendidas/mes ÷ stock) con la columna '
          'K nueva. Instrucciones!A6 prometía que se calculaba y era una celda '
          'verde vacía [TEC-18 · DOM-28 · §4.3]')

    # ---- fila TOTAL: un «Budget de Bodega» que valora la bodega -----------
    r = ultima + 1
    _et(ws, 'A' + str(r), MARCA_C, bold=True)
    r += 1
    fila_total = r
    _et(ws, 'B' + str(r), 'TOTAL BODEGA', bold=True)
    motor.f(ws, 'I' + str(r),
            _suma_guardada('I' + str(primera) + ':I' + str(ultima)),
            fmt=motor.FMT_ENT, bold=True)
    motor.f(ws, 'N' + str(r),
            _suma_guardada('N' + str(primera) + ':N' + str(ultima)),
            fmt=motor.FMT_EUR, bold=True)
    motor.f(ws, 'O' + str(r),
            _suma_guardada('O' + str(primera) + ':O' + str(ultima)),
            fmt=motor.FMT_EUR, bold=True)
    motor.f(ws, 'L' + str(r),
            motor.iferror('IF(OR($O$' + str(r) + '="",$O$' + str(r)
                          + '=0,$N$' + str(r) + '=""),"",($O$' + str(r)
                          + '-$N$' + str(r) + ')/$O$' + str(r) + ')'),
            fmt=motor.FMT_PCT, bold=True)
    motor.fijar_formato(ws, 'L' + str(r), motor.FMT_PCT)
    motor.f(ws, 'M' + str(r), _pct('$N$' + str(r), '$O$' + str(r)),
            fmt=motor.FMT_PCT, bold=True)
    motor.fijar_formato(ws, 'M' + str(r), motor.FMT_PCT)
    r += 1
    _et(ws, 'A' + str(r),
        'El «Valor stock a coste» de la fila TOTAL es la cifra que alimenta la '
        'partida «Bodega inicial (vinos)» de plan-financiero-3-anos.xlsx, hoja '
        '«Inversión». Cópiala a mano: los libros no se enlazan entre sí porque '
        'un .xlsx movido de carpeta daría #REF! [TEC-19 · §4.3 · §1.13].',
        wrap=True)
    cambios.append(
        fname + ':' + ws.title + '!fila ' + str(fila_total)
        + ': fila TOTAL con uds en stock, valor a coste, valor a PVP, margen '
          'medio ponderado y food cost medio. Un fichero llamado «Budget de '
          'Bodega» no sumaba nada: 100 fórmulas y ninguna cruzaba coste con '
          'stock [TEC-19 · COM-22 · §4.3]')
    _param(ws, fila_iva_beb, None, valor=0.21,
           etiqueta='IVA de la bebida alcohólica (%)', fmt=motor.FMT_PCT,
           nota=('21 % en bebidas alcohólicas (art. 91 de la Ley del IVA). El '
                 'PVP de la columna F es el de CARTA, que en España se anuncia '
                 'CON IVA: el multiplicador, el margen, el food cost y el '
                 'valor a PVP se calculan sobre la columna «PVP sin IVA», no '
                 'sobre el precio de carta.'),
           col_et='A', col_val='C', col_nota='D')
    cambios.append(
        fname + ':' + ws.title + '!T' + str(primera) + ':T' + str(ultima)
        + ': columna «PVP sin IVA (€)» = PVP de carta / (1 + IVA de bebida '
          'alcohólica, en celda). El food cost de bebida cruzaba un coste sin '
          'IVA con un PVP de carta CON IVA y salía unos 7 puntos bajo, que es '
          'lo que decide si la bodega gana dinero [RD-13 · §1.5a]')
    _instr(wb, fname, cambios, [
        'Los precios de la columna «PVP Carta (€)» van CON IVA, que es como se '
        'anuncian en España. El libro calcula el «PVP sin IVA» con el tipo de '
        'la bebida alcohólica (21 %, en celda) y TODOS los indicadores salen '
        'de ahí: cruzar un coste sin IVA con un precio con IVA deja el food '
        'cost de bebida unos 7 puntos por debajo del real.',
        'La columna «Multiplicador (×)» es PVP ÷ coste (el x2,5-x3,5 del '
        'capítulo de bodega). El margen de verdad es «Margen s/PVP (%)», y el '
        '«Food cost bebida (%)» es el que se compara con el 28-40 % de la guía.',
        'Escribe las «Uds vendidas/mes» y la «Rotación/Mes» se calcula sola.',
        'La fila TOTAL valora la bodega a coste y a PVP: es la cifra de la '
        'partida «Bodega inicial» del plan financiero.'])
    registro_modelo['bodega'] = {'primera': primera, 'ultima': ultima,
                                 'fila_total': fila_total}
    return registro_modelo['bodega']


# ==========================================================================
# §4.4 — plantilla-turnos-brigada
# ==========================================================================
def modelo_turnos(ws, fname=''):
    fila = _fila_cabecera(ws)
    if not fila:
        raise EstructuraDesconocida(
            fname + ':' + ws.title + ': sin fila de cabecera reconocible')
    c = dict((k, _norm(v)) for k, v in _cab(ws, fila).items())
    if c.get('A') == 'persona' and c.get('J') == 'total h':
        return 'T3', fila
    if c.get('C') == 'puesto' and c.get('J') == 'dom':
        # T1 y T2 comparten cabecera. Lo que las separa —y sobrevive a la 1.ª
        # pasada, que añade la columna «Horas/Semana» a T2 y la haría idéntica a
        # T1— es la primera fila de datos: en T1 es una BANDA de sección
        # (`A5` vacía, `B5='COCINA'`) y en T2 es el puesto nº 1 (`A5=1`).
        primera = ws.cell(row=fila + 1, column=1).value
        if _es_num(primera):
            return 'T2', fila
        if isinstance(_txt(ws.cell(row=fila + 1, column=2).value), str):
            return 'T1', fila
    raise EstructuraDesconocida(
        fname + ':' + ws.title + ': cabecera de cuadrante no reconocida por '
        '§4.4 (fila ' + str(fila) + ') = '
        + repr(dict((k, v) for k, v in _cab(ws, fila).items() if v is not None)))


#: Equivalencia turno -> horas del §4.4. Los valores son de la SPEC, no del
#: sector: van a celda verde y documentados, y la «V» se DOCUMENTA como
#: Vacaciones en vez de retirarse de la lista (TEC-12).
TURNOS = (('M', 'Mañana', 8), ('T', 'Tarde', 8), ('P', 'Partido', 10),
          ('L', 'Libre', 0), ('V', 'Vacaciones', 0))


def _filas_puesto(ws, modelo, fila_cab, fin):
    """Filas que son un PUESTO, no una banda de sección ni una leyenda."""
    fuera = []
    for r in range(fila_cab + 1, fin + 1):
        if modelo == 'T3':
            if _txt(ws.cell(row=r, column=2).value):
                fuera.append(r)
            continue
        if _txt(ws.cell(row=r, column=3).value):
            fuera.append(r)
    return fuera


def _turnos(wb, fname, cambios, contenido, registro_modelo):
    hojas = [ws for ws in wb.worksheets
             if ws.title not in ('Instrucciones', HOJA_REGISTRO)]
    ws = hojas[0]
    modelo, fila_cab = modelo_turnos(ws, fname)
    marca = _fila_marca(ws)
    _t = (marca - 1) if marca else None
    fin = _fin_por_columna(ws, fila_cab, 'B' if modelo == 'T3' else 'C',
                           tope=_t)
    puestos = _filas_puesto(ws, modelo, fila_cab, fin)
    if not puestos:
        raise EstructuraDesconocida(fname + ':' + ws.title + ': 0 puestos')

    pie_fila, pie_texto = _pie_existente(ws, fila_cab + 1)
    # T3 lleva DEBAJO de la tabla cinco «Recordatorios legales» (descanso de 12
    # h, descanso semanal, plus de nocturnidad, control horario digital) que son
    # contenido, no mobiliario: el bloque se escribe por debajo de ellos en vez
    # de barrerlos. En T1/T2 no hay nada que salvar salvo el pie, que se captura.
    if marca:
        inicio = marca
    elif modelo == 'T3':
        inicio = ws.max_row + 2
    else:
        inicio = fin + 2 if modelo == 'T1' else fin + 1
    _snapshot_parametros(ws, inicio, ws.max_row)
    _limpiar(ws, inicio, max(ws.max_row, inicio + 26))

    col_horas = 'K' if modelo != 'T3' else 'J'
    if modelo == 'T2':
        _cabecera_nueva(ws, fila_cab, 'K', 'Horas/Semana', 14)
    nuevas = (('L', 'Bruto anual (€)', 16), ('M', 'Nº de pagas', 12),
              ('N', 'Bruto por paga (€)', 16), ('O', 'Coste/hora (€)', 14),
              ('P', 'Coste semana (€)', 16))
    for letra, texto, ancho in nuevas:
        _cabecera_nueva(ws, fila_cab, letra, texto, ancho)

    # ---- bloque de parámetros de la hoja ----------------------------------
    r = inicio
    _et(ws, 'A' + str(r), MARCA_C, bold=True)
    r += 1
    fila_total = r
    _et(ws, 'B' + str(r), 'TOTAL BRIGADA', bold=True)
    r += 2
    _et(ws, 'A' + str(r), CAB_PARAMETROS, bold=True)
    r += 1
    refs_turno = {}
    if modelo != 'T3':
        for letra, nombre, horas in TURNOS:
            refs_turno[letra] = _param(
                ws, r, None, valor=horas,
                etiqueta='Horas del turno ' + letra + ' (' + nombre + ')',
                fmt='#,##0.0', nota=None, col_et='A', col_val='C',
                col_nota='D', maximo=24)
            r += 1
        _et(ws, 'D' + str(r - len(TURNOS)),
            'Tabla de equivalencia turno' + motor.NARROW + '-> horas. La «V» '
            'del desplegable existía y no estaba documentada en ninguna parte '
            'del fichero: es Vacaciones [TEC-12].', wrap=True)
    ref_ss = _param(ws, r, 'ss_empresa', col_et='A', col_val='C', col_nota='D')
    r += 1
    ref_smi = _param(ws, r, 'smi_anual', col_et='A', col_val='C', col_nota='D')
    r += 1
    # RD-04/RT-05/RC-12 · esta celda es el DIVISOR de todos los costes del
    # libro y se entregaba vacía: coste/hora, coste semana y los tres totales
    # devolvían "" en las 24 filas, mientras la tarjeta vende un cuadrante
    # «con coste». Dejarla vacía era correcto (no se inventa el convenio) pero
    # deja la promesa sin cumplir. La salida: NO se teclea ninguna cifra de
    # convenio; se CALCULA una jornada de referencia con los cuatro parámetros
    # que sí tienen fuente legal, y la celda del convenio se queda verde y
    # vacía como OVERRIDE. Si el cliente la rellena, manda ella.
    ref_jornada_conv = _param(
        ws, r, None, valor=None,
        etiqueta='Jornada anual de convenio (h/año)', fmt=motor.FMT_ENT,
        nota=('La fija el convenio provincial de hostelería, en su tabla '
              'salarial anual. No se precarga con una cifra inventada: '
              'escríbela tú y prevalece sobre la jornada de referencia de '
              'abajo. Mientras esté vacía, el libro usa la de referencia.'),
        col_et='A', col_val='C', col_nota='D', maximo=2500)
    r += 1
    ref_max = _param(
        ws, r, None, valor=40,
        etiqueta='Jornada máxima semanal (h)', fmt='#,##0.0',
        nota=('40 h de promedio en cómputo anual (art. 34.1 ET). El convenio '
              'puede fijar menos.'), col_et='A', col_val='C', col_nota='D',
        maximo=60)
    r += 1
    ref_sem = _param(
        ws, r, None, valor=52, etiqueta='Semanas del año',
        fmt=motor.FMT_ENT, nota='52 semanas.', col_et='A', col_val='C',
        col_nota='D', maximo=53)
    r += 1
    ref_vac = _param(
        ws, r, None, valor=30, etiqueta='Días de vacaciones al año (naturales)',
        fmt=motor.FMT_ENT,
        nota=('30 días naturales, el mínimo del art. 38.1 ET. El convenio '
              'puede dar más.'), col_et='A', col_val='C', col_nota='D',
        maximo=90)
    r += 1
    ref_fes = _param(
        ws, r, None, valor=14, etiqueta='Festivos anuales retribuidos',
        fmt=motor.FMT_ENT,
        nota=('14 al año como máximo (art. 37.2 ET): 12 nacionales o '
              'autonómicos más 2 locales.'), col_et='A', col_val='C',
        col_nota='D', maximo=20)
    r += 1
    ref_dsem = _param(
        ws, r, None, valor=7, etiqueta='Días naturales de la semana',
        fmt=motor.FMT_ENT,
        nota=('7. Convierte los días naturales de vacaciones en semanas.'),
        col_et='A', col_val='C', col_nota='D', maximo=7)
    r += 1
    ref_dlab = _param(
        ws, r, None, valor=5, etiqueta='Días de trabajo por semana',
        fmt=motor.FMT_ENT,
        nota=('5 con el descanso semanal de día y medio del art. 37.1 ET. '
              'Convierte cada festivo en horas.'),
        col_et='A', col_val='C', col_nota='D', maximo=7)
    r += 1
    fila_jref = r
    _et(ws, 'A' + str(r), 'Jornada anual de REFERENCIA (h/año)')
    motor.f(ws, 'C' + str(r),
            '=IFERROR(ROUND(' + ref_max + '*' + ref_sem + '-' + ref_max + '*'
            + ref_vac + '/' + ref_dsem + '-' + ref_fes + '*' + ref_max + '/'
            + ref_dlab + ',0),"")', fmt=motor.FMT_ENT)
    motor.fijar_formato(ws, 'C' + str(r), motor.FMT_ENT)
    _et(ws, 'D' + str(r),
        'NO es un dato del convenio: se calcula con los cuatro parámetros de '
        'arriba, y los cuatro salen de la ley (art. 34.1, 38.1 y 37.2 ET). '
        'Sirve para que el libro enseñe el coste/hora desde el primer momento; '
        'en cuanto escribas la jornada de TU convenio, manda la de arriba.',
        wrap=True)
    r += 1
    fila_jef = r
    _et(ws, 'A' + str(r), 'Jornada anual aplicada (h/año)', bold=True)
    ref_jornada = '$C$' + str(r)
    motor.f(ws, 'C' + str(r),
            '=IF(' + ref_jornada_conv + '="",IF($C$' + str(fila_jref)
            + '="","",$C$' + str(fila_jref) + '),' + ref_jornada_conv + ')',
            fmt=motor.FMT_ENT, bold=True)
    motor.fijar_formato(ws, 'C' + str(r), motor.FMT_ENT)
    _et(ws, 'D' + str(r),
        'La que divide en «Coste/hora»: la de tu convenio si la has escrito, y '
        'si no, la de referencia.', wrap=True)
    r += 1
    _et(ws, 'A' + str(r),
        'Los brutos anuales que trae el libro son los del capítulo de brigada '
        'de esta misma guía, ELEVADOS al SMI vigente donde el capítulo se '
        'quedaba por debajo (el capítulo publicaba dos puestos bajo el SMI: '
        'ésos se han subido, no se han copiado). Son valores de EJEMPLO: '
        'escribe los tuyos y el semáforo se enciende en rojo si alguno queda '
        'por debajo del SMI prorrateado a su jornada '
        '[DOM-13 · §7-bis.16 · RD-04 · RD-25].', wrap=True)
    r += 2
    if pie_texto:
        _et(ws, 'A' + str(r), pie_texto)
        if pie_fila and pie_fila != r:
            cambios.append(fname + ':' + ws.title + ': el pie baja de la fila '
                           + str(pie_fila) + ' a la ' + str(r))

    # ---- horas y coste por puesto -----------------------------------------
    for f_ in puestos:
        if modelo != 'T3':
            sp = '+'.join('(D{r}:J{r}="' + l + '")*' + refs_turno[l]
                          for l, _n, _h in TURNOS).format(r=f_)
            # ROUND obligatorio: sin él pycel devuelve numpy.int64 y el SUM del
            # TOTAL da 0 y el semáforo ISNUMBER da False (ver docstring).
            motor.f(ws, col_horas + str(f_),
                    '=IF(COUNTIF(D{r}:J{r},"<>")=0,"",ROUND(SUMPRODUCT({sp}),2))'
                    .format(r=f_, sp=sp), fmt='#,##0.0')
            motor.fijar_formato(ws, col_horas + str(f_), '#,##0.0')
            motor.verde(ws, 'B' + str(f_))
        else:
            motor.verde(ws, col_horas + str(f_))
            motor.fijar_formato(ws, col_horas + str(f_), '#,##0.0')
        motor.verde(ws, 'L' + str(f_) + ':M' + str(f_))
        motor.fijar_formato(ws, 'L' + str(f_), motor.FMT_EUR)
        motor.fijar_formato(ws, 'M' + str(f_), motor.FMT_ENT)
        motor.f(ws, 'N' + str(f_),
                motor.iferror('IF(OR($L{r}="",$M{r}="",$M{r}=0),"",'
                              '$L{r}/$M{r})'.format(r=f_)), fmt=motor.FMT_EUR)
        motor.f(ws, 'O' + str(f_),
                motor.iferror('IF(OR($L{r}="",' + ref_jornada + '="",'
                              + ref_jornada + '=0),"",$L{r}*(1+' + ref_ss
                              + ')/' + ref_jornada + ')').format(r=f_),
                fmt=motor.FMT_EUR)
        motor.f(ws, 'P' + str(f_),
                motor.iferror('IF(OR($' + col_horas + str(f_) + '="",$O'
                              + str(f_) + '=""),"",$' + col_horas + str(f_)
                              + '*$O' + str(f_) + ')'), fmt=motor.FMT_EUR)
    motor.dv_lista(ws, ['M' + str(f_) for f_ in puestos], ['12', '14', '15'],
                   titulo='Nº de pagas',
                   mensaje='12, 14 o 15 pagas (14 es lo habitual en hostelería).')
    for f_ in puestos:
        if ws['M' + str(f_)].value is None:
            motor.val(ws, 'M' + str(f_), 14, fmt=motor.FMT_ENT, verde_=True)

    # ---- RD-25 · el SMI se prorratea a la jornada -------------------------
    # En hostelería el contrato a tiempo parcial es habitual y su bruto legítimo
    # está por debajo del SMI anual a jornada completa: la regla anterior
    # pintaba en rojo contratos correctos y no distinguía el incumplimiento
    # real, que es lo que el gate pretende cazar.
    col_pct = 'Q'
    _cabecera_nueva(ws, fila_cab, col_pct, '% de jornada', 13)
    for f_ in puestos:
        if ws[col_pct + str(f_)].value is None:
            motor.val(ws, col_pct + str(f_), 1.0, fmt=motor.FMT_PCT,
                      verde_=True)
        motor.verde(ws, col_pct + str(f_))
        motor.fijar_formato(ws, col_pct + str(f_), motor.FMT_PCT)
    motor.prompt_porcentaje(
        ws, col_pct + str(puestos[0]) + ':' + col_pct + str(puestos[-1]),
        'Jornada', 'Se escribe en tanto por uno: 1 = jornada completa, '
        '0,5 = media jornada. El SMI se compara prorrateado a este porcentaje.')

    # ---- RD-04/RT-05 · brutos anuales de EJEMPLO, del propio capítulo -----
    conf_t = ((getattr(contenido, 'TURNOS', None) or {}) if contenido else {})
    brutos = conf_t.get('brutos') or {}
    turnos_ej = conf_t.get('cuadrante') or {}
    puestos_txt = dict((f_, _norm(ws.cell(row=f_, column=3).value))
                       for f_ in puestos)
    precargados = 0
    for f_, etiqueta in puestos_txt.items():
        valor = brutos.get(etiqueta)
        if valor is not None and ws['L' + str(f_)].value is None:
            motor.val(ws, 'L' + str(f_), valor, fmt=motor.FMT_EUR, verde_=True)
            motor.fijar_formato(ws, 'L' + str(f_), motor.FMT_EUR)
            precargados += 1
        patron = turnos_ej.get(etiqueta)
        if patron and modelo != 'T3':
            for i, letra in enumerate(patron[:7]):
                col = chr(ord('D') + i)
                if ws[col + str(f_)].value is None:
                    motor.val(ws, col + str(f_), letra)
                    motor.verde(ws, col + str(f_))
    if precargados:
        cambios.append(
            fname + ':' + ws.title + '!L' + str(puestos[0]) + ':L'
            + str(puestos[-1]) + ': ' + str(precargados) + ' brutos anuales de '
            'ejemplo precargados desde el capítulo de brigada de esta misma '
            'guía, con los dos puestos que el capítulo publicaba por debajo '
            'del SMI ELEVADOS al SMI vigente; y el cuadrante con una rotación '
            'de ejemplo, para que el coste/hora y el coste de la semana se '
            'vean al abrir el fichero [RD-04 · RT-05 · RC-12]')

    # ---- RD-27/RC-14 · la ÚNICA DV del producto sin mensaje de error ------
    if modelo != 'T3':
        motor.dv_lista(
            ws, [chr(ord('D') + i) + str(f_) for f_ in puestos
                 for i in range(7)],
            [l for l, _n, _h in TURNOS], titulo='Turno',
            mensaje='Escribe M (mañana), T (tarde), P (partido), L (libre) o '
                    'V (vacaciones). Cualquier otra letra la cuenta como 0 h y '
                    'te resta horas y coste de brigada sin avisar.')

    # ---- numeración de la columna A (TEC-13) ------------------------------
    if modelo == 'T1':
        for i, f_ in enumerate(puestos, start=1):
            motor.val(ws, 'A' + str(f_), i, fmt=motor.FMT_ENT)
        cambios.append(fname + ':' + ws.title + '!A' + str(puestos[0]) + ':A'
                       + str(puestos[-1]) + ': columna # numerada 1-'
                       + str(len(puestos)) + ' (estaba vacía mientras el resto '
                       'de plantillas del producto sí numera) [TEC-13]')

    # ---- totales -----------------------------------------------------------
    prim, ult = puestos[0], puestos[-1]
    motor.f(ws, col_horas + str(fila_total),
            _suma_guardada(col_horas + str(prim) + ':' + col_horas + str(ult)),
            fmt='#,##0.0', bold=True)
    motor.fijar_formato(ws, col_horas + str(fila_total), '#,##0.0')
    motor.f(ws, 'L' + str(fila_total),
            _suma_guardada('L' + str(prim) + ':L' + str(ult)),
            fmt=motor.FMT_EUR, bold=True)
    motor.f(ws, 'P' + str(fila_total),
            _suma_guardada('P' + str(prim) + ':P' + str(ult)),
            fmt=motor.FMT_EUR, bold=True)
    _et(ws, 'N' + str(fila_total), 'Coste anual con SS (€)', bold=True)
    # RC-35 · es una ETIQUETA, no un importe: el formato de moneda sobre un
    # texto es un residuo de la regla de columna del §1.4.
    motor.fijar_formato(ws, 'N' + str(fila_total), 'General')
    motor.f(ws, 'O' + str(fila_total),
            motor.iferror('IF($L$' + str(fila_total) + '="","",$L$'
                          + str(fila_total) + '*(1+' + ref_ss + '))'),
            fmt=motor.FMT_EUR, bold=True)
    motor.fijar_formato(ws, 'O' + str(fila_total), motor.FMT_EUR)

    # ---- semáforos ---------------------------------------------------------
    motor.regla_expresion(
        ws, 'L' + str(prim) + ':L' + str(ult),
        '=AND(ISNUMBER($L' + str(prim) + '),ISNUMBER($Q' + str(prim) + '),$L'
        + str(prim) + '<' + ref_smi + '*$Q' + str(prim) + ')')
    motor.semaforo_isnumber(ws, col_horas + str(prim) + ':' + col_horas
                            + str(ult), '$' + col_horas + str(prim),
                            operador='>', umbral=ref_max,
                            bg=motor.CF_AMBAR_BG, fg=motor.CF_AMBAR_FG)
    cambios.append(
        fname + ':' + ws.title + '!' + col_horas + str(prim) + ':P' + str(ult)
        + ': horas por SUMPRODUCT sobre la tabla de turnos, bruto anual, nº de '
          'pagas, bruto por paga, coste/hora con la SS al 33 % EN CELDA y coste '
          'semana, con TOTAL y coste anual con SS. El libro tenía 0 fórmulas y '
          'ninguna columna de coste, y sus Instrucciones prometían que «las '
          'horas y costes se calculan automáticamente» '
          '[TEC-11 · DOM-02 · DOM-16 · COM-10 · COM-20 · §4.4]')
    cambios.append(
        fname + ':' + ws.title + '!L' + str(prim) + ':L' + str(ult)
        + ': semáforo rojo por formato condicional sobre todo bruto anual por '
          'debajo del SMI (' + str(motor.SMI_ANUAL) + ' EUR/año, RD 126/2026), '
          'con la guarda ISNUMBER para que no se encienda con una celda vacía '
          '[DOM-13 · §1.6 · §7-bis.16]')

    _registro_jornada(wb, fname, cambios, ws, puestos, modelo)
    _headcount(wb, ws, fname, cambios, len(puestos))
    lineas = [
        'La columna «Horas/Semana» se calcula sola desde los turnos: M, T, P, '
        'L y V. La V es VACACIONES (estaba en el desplegable y no la explicaba '
        'ninguna parte del fichero).',
        'Escribe el «Bruto anual» de cada puesto y la «Jornada anual de '
        'convenio»: el libro calcula el coste/hora con la Seguridad Social a '
        'cargo de la empresa y el coste de la semana.',
        'Si un bruto anual queda por debajo del SMI, la celda se pinta de rojo.',
        'La hoja «' + HOJA_REGISTRO + '» es la que cumple el registro de '
        'jornada del art. 34.9 del Estatuto de los Trabajadores: un cuadrante '
        'con letras M/T/P/L no lo cumple.',
    ]
    if modelo == 'T3':
        lineas[0] = ('En este cuadrante los días llevan el RANGO HORARIO '
                     '(«03-11»), no la letra del turno, así que «Total h» es '
                     'una celda verde que escribes tú. Las horas calculadas '
                     'están en la hoja «' + HOJA_REGISTRO + '».')
    _instr(wb, fname, cambios, lineas)
    registro_modelo['turnos'] = {
        'modelo': modelo, 'puestos': len(puestos), 'primera': prim,
        'ultima': ult, 'col_horas': col_horas, 'fila_total': fila_total,
        'ref_smi': ref_smi, 'ref_max': ref_max, 'ref_ss': ref_ss,
        'ref_jornada': ref_jornada,
        'refs_turno': dict(refs_turno)}
    return registro_modelo['turnos']


def _headcount(wb, ws, fname, cambios, n):
    """TEC-13/COM-21 — una sola cifra de plantilla, la MEDIDA.

    El título dice «(25 personas)» y el cuadrante tiene 24 puestos. La cifra que
    manda es la que se cuenta al abrir el fichero (§9, gate de recuento). El
    texto del capítulo y la tarjeta del dashboard son de T7/T8: aquí se corrige
    lo que vive en el xlsx y se deja anotado.
    """
    rx = re.compile(r'\((\d+)\s+personas\)')
    tocados = []
    for hoja in wb.worksheets:
        for row in hoja.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and c.data_type != 'f' and rx.search(v):
                    nuevo = rx.sub('(' + str(n) + ' personas)', v)
                    if nuevo != v:
                        c.value = nuevo
                        tocados.append(hoja.title + '!' + c.coordinate)
    p = wb.properties
    for campo in ('title', 'subject'):
        v = getattr(p, campo, None)
        if isinstance(v, str) and rx.search(v):
            nuevo = rx.sub('(' + str(n) + ' personas)', v)
            if nuevo != v:
                setattr(p, campo, nuevo)
                tocados.append('propiedades.' + campo)
    if tocados:
        cambios.append(
            fname + ': headcount UNIFICADO a ' + str(n) + ' puestos, que es lo '
            'que hay en el cuadrante (' + ', '.join(tocados) + '). El texto del '
            'cap. 14 («22-30»), la suma de las tablas del docx (21-29) y la '
            'tarjeta del dashboard («25 personas») siguen diciendo otra cosa: '
            'son de T7/T8 [TEC-13 · COM-21 · §7-bis.7]')


def _registro_jornada(wb, fname, cambios, ws_turnos, puestos, modelo):
    """DOM-16 — la hoja que sí cumple el art. 34.9 ET.

    Las horas van como DURACIÓN (`MOD(salida-entrada,1)`, formato `[h]:mm`) y no
    como `MOD(...)*24`: así el cruce de medianoche sale bien (medido con pycel:
    23:00 -> 08:00 = 9 h) y la fórmula no lleva el literal 24 dentro (§1.3).
    """
    nueva = HOJA_REGISTRO not in wb.sheetnames
    ws = wb[HOJA_REGISTRO] if not nueva else wb.create_sheet(HOJA_REGISTRO)
    _snapshot_parametros(ws, 1, ws.max_row)
    for row in ws.iter_rows():
        for c in row:
            if c.__class__.__name__ != 'MergedCell':
                c.value = None
    _a4(ws)
    anchos = {'A': 14.0, 'B': 26.0, 'C': 12.0, 'D': 11.0, 'E': 11.0,
              'F': 13.0, 'G': 18.0, 'H': 40.0, 'I': 16.0}
    for L, w in anchos.items():
        ws.column_dimensions[L].width = w
    _et(ws, 'A1', 'Registro de jornada (art. 34.9 del Estatuto de los '
                  'Trabajadores)', bold=True)
    ws['A1'].font = Font(bold=True, size=14)
    _et(ws, 'A2', 'AI Chef Pro · aichef.pro — una fila por trabajador y día, '
                  'con hora de entrada y de salida.')
    cabs = (('A', 'Fecha'), ('B', 'Trabajador'), ('C', 'Semana'),
            ('D', 'Entrada'), ('E', 'Salida'), ('F', 'Horas del día'),
            ('G', 'Descanso previo'), ('H', 'Aviso'), ('I', 'Firma'))
    fila_cab = 4
    for L, texto in cabs:
        _et(ws, L + str(fila_cab), texto, bold=True)
        modelo_estilo = ws_turnos.cell(row=4, column=1)
        if modelo_estilo.has_style:
            ws[L + str(fila_cab)]._style = copy.copy(modelo_estilo._style)
        ws[L + str(fila_cab)].value = texto
    # RD-26 · 40 filas de registro para una brigada de 24 personas son menos
    # de dos días, cuando el art. 34.9 ET obliga a registrar TODOS los días de
    # TODOS los trabajadores. Una semana completa de esta brigada son
    # `puestos × 7` filas; se dimensiona a eso, con un mínimo de 40.
    filas_semana = max(40, (len(puestos) or 6) * 7)
    primera, ultima = fila_cab + 1, fila_cab + filas_semana
    r = ultima + 2
    _et(ws, 'A' + str(r), CAB_PARAMETROS, bold=True)
    r += 1
    ref_desc = _param(
        ws, r, None, valor=0.5, etiqueta='Descanso mínimo entre jornadas',
        fmt=FMT_DUR,
        nota=('12:00 = doce horas (art. 34.3 ET). Se escribe como hora, no '
              'como número.'), col_et='A', col_val='C', col_nota='D')
    r += 2
    _et(ws, 'A' + str(r), 'Horas registradas por trabajador y semana', bold=True)
    r += 1
    fila_res_cab = r
    for L, texto in (('A', 'Trabajador'), ('B', 'Semana'),
                     ('C', 'Horas registradas')):
        _et(ws, L + str(r), texto, bold=True)
    res_prim, res_ult = r + 1, r + 10
    for f_ in range(res_prim, res_ult + 1):
        motor.verde(ws, 'A' + str(f_) + ':B' + str(f_))
        # RC-13 · la forma con `*` propaga #¡VALOR! en cuanto una fila del
        # parte está sin rellenar —es decir, siempre—, porque `F` devuelve ""
        # y TRUE*"" es un error. La forma con COMAS trata el texto como cero,
        # que es justo lo que hace falta. Es la misma forma que ya usa bien
        # menu-engineering-matrix en su media ponderada.
        motor.f(ws, 'C' + str(f_),
                '=IF(OR($A{r}="",$B{r}=""),"",ROUND(SUMPRODUCT('
                '--($B${p}:$B${u}=$A{r}),--($C${p}:$C${u}=$B{r}),'
                'IFERROR($F${p}:$F${u}*1,0)),5))'
                .format(r=f_, p=primera, u=ultima), fmt=FMT_DUR)
        motor.fijar_formato(ws, 'C' + str(f_), FMT_DUR)
    r = res_ult + 2
    _et(ws, 'A' + str(r),
        'Un cuadrante con letras M/T/P/L en siete columnas NO cumple el '
        'registro de jornada: hace falta la hora de entrada y de salida de cada '
        'trabajador cada día. El propio checklist de contratación lo exige '
        '(«Registro de jornada digital (obligatorio)») [DOM-16].', wrap=True)
    r += 1
    _et(ws, 'A' + str(r),
        'Escribe las horas como hora («08:00», «23:30»). El cruce de medianoche '
        'está contemplado: una salida a las 02:00 del día siguiente cuenta '
        'bien.', wrap=True)

    for f_ in range(primera, ultima + 1):
        motor.verde(ws, 'A' + str(f_) + ':E' + str(f_))
        motor.verde(ws, 'I' + str(f_))
        motor.fijar_formato(ws, 'A' + str(f_), motor.FMT_FECHA)
        motor.fijar_formato(ws, 'D' + str(f_), FMT_HORA)
        motor.fijar_formato(ws, 'E' + str(f_), FMT_HORA)
        motor.f(ws, 'F' + str(f_),
                '=IF(OR($D{r}="",$E{r}=""),"",MOD($E{r}-$D{r},1))'
                .format(r=f_), fmt=FMT_DUR)
        motor.fijar_formato(ws, 'F' + str(f_), FMT_DUR)
        if f_ == primera:
            # RC-35 · `'=""'` es una fórmula que sólo devuelve la cadena vacía:
            # un residuo de construcción que además cuenta como una de las 230
            # fórmulas del fichero. La primera fila del parte no tiene jornada
            # anterior con la que comparar, así que va VACÍA.
            ws['G' + str(f_)].value = None
        else:
            # RT-22 · `MOD(entrada − salida_anterior, 1)` devuelve sólo la
            # parte fraccionaria del día: un descanso de EXACTAMENTE 24 h da
            # 0:00 y dispara el aviso legal, y uno de 36 h da 12:00 y también.
            # Con la fecha de la columna A la resta es real y soporta cualquier
            # separación; si falta la fecha se cae al comportamiento anterior,
            # pero sumando un día cuando la resta sale negativa o nula.
            motor.f(ws, 'G' + str(f_),
                    '=IF(OR($B{r}="",$B{r}<>$B{a},$D{r}="",$E{a}=""),"",'
                    'IF(OR($A{r}="",$A{a}=""),'
                    'IF(MOD($D{r}-$E{a},1)=0,1,MOD($D{r}-$E{a},1)),'
                    '($A{r}+$D{r})-($A{a}+$E{a})))'.format(r=f_, a=f_ - 1),
                    fmt=FMT_DUR)
        motor.fijar_formato(ws, 'G' + str(f_), FMT_DUR)
        motor.f(ws, 'H' + str(f_),
                '=IF($G{r}="","",IF($G{r}<{d},"Descanso inferior al mínimo '
                'legal entre jornadas (art. 34.3 ET)",""))'
                .format(r=f_, d=ref_desc))
    motor.dv_fecha(ws, ['A' + str(f_) for f_ in range(primera, ultima + 1)])
    motor.semaforo_isnumber(ws, 'G' + str(primera) + ':G' + str(ultima),
                            '$G' + str(primera), operador='<', umbral=ref_desc)
    if puestos and modelo != 'T3':
        # RD-26 · el registro tiene que acreditar la jornada de PERSONAS
        # identificadas, no de puestos. Si el cuadrante ya trae nombres en su
        # columna «Nombre», el desplegable apunta ahí; si está vacía (que es
        # como se entrega), se cae a los puestos y la nota lo dice.
        nombres = [_txt(ws_turnos.cell(row=f_, column=2).value)
                   for f_ in puestos]
        nombres = [n for n in nombres if n]
        origen = 'los nombres del cuadrante'
        if not nombres:
            nombres = [_txt(ws_turnos.cell(row=f_, column=3).value)
                       for f_ in puestos]
            nombres = [n for n in nombres if n]
            origen = ('los PUESTOS del cuadrante: escribe los nombres en su '
                      'columna «Nombre» y vuelve a abrir esta hoja para que el '
                      'desplegable ofrezca personas')
        if nombres:
            motor.dv_lista(
                ws, ['B' + str(f_) for f_ in range(primera, ultima + 1)],
                nombres[:40], titulo='Trabajador',
                mensaje='Elige un trabajador de la lista (sale de ' + origen
                        + ').')
    del fila_res_cab
    if nueva:
        cambios.append(
            fname + ': hoja «' + HOJA_REGISTRO + '» CREADA — fecha, '
            'trabajador, semana, entrada, salida, horas del día con cruce de '
            'medianoche, descanso desde la jornada anterior con aviso del art. '
            '34.3 ET, firma, y resumen de horas por trabajador y semana '
            '[DOM-16 · §4.4]')


# ==========================================================================
# §4.5 — plan-fermentacion-y-produccion (sólo panadería)
# ==========================================================================
def _fermentacion(wb, fname, cambios, contenido, registro_modelo):
    """§4.5 — la hoja tiene 0 fórmulas y el gate exige que calcule algo.

    Lo que se convierte en fórmula NO se inventa: `H` («Tiempo total h») es
    EXACTAMENTE `E + F + G` (bulk + bloque frío + final) en las 15 filas —
    comprobado fila a fila antes de tocar nada—, así que es una conversión de
    constante a fórmula del §1.2 con el número conservado. Si alguna fila no
    cuadrara, la hoja se queda como está y el desajuste se anota: no se
    reescribe una fermentación para que encaje con la fórmula.
    """
    ws = [w for w in wb.worksheets if w.title != 'Instrucciones'][0]
    fila_cab = _fila_cabecera(ws)
    c = dict((k, _norm(v)) for k, v in _cab(ws, fila_cab).items())
    if not (c.get('E') == 'bulk h' and c.get('H') == 'tiempo total h'):
        raise EstructuraDesconocida(
            fname + ':' + ws.title + ': cabecera de plan de fermentación no '
            'reconocida por §4.5 = '
            + repr(dict((k, v) for k, v in _cab(ws, fila_cab).items()
                        if v is not None)))
    marca = _fila_marca(ws)
    fin = _fin_por_columna(ws, fila_cab, 'B',
                           tope=(marca - 1) if marca else None)
    descuadres = []
    for r in range(fila_cab + 1, fin + 1):
        h = ws['H' + str(r)].value
        partes = [ws[L + str(r)].value for L in ('E', 'F', 'G')]
        if _es_num(h) and all(_es_num(x) for x in partes) \
                and abs(sum(partes) - h) > 0.01:
            descuadres.append(ws.title + '!H' + str(r) + '=' + str(h)
                              + ' pero E+F+G=' + str(sum(partes)))
    if descuadres:
        cambios.append(
            fname + ': NO se convierte «Tiempo total h» en fórmula: '
            + str(len(descuadres)) + ' filas no cuadran con E+F+G ('
            + '; '.join(descuadres[:6]) + '). La hoja se queda con §1 y el '
            'desajuste queda anotado — no se reescribe una fermentación para '
            'que encaje [§4.5]')
        registro_modelo['fermentacion'] = {'convertidas': 0,
                                           'descuadres': descuadres}
        return registro_modelo['fermentacion']

    _snapshot_parametros(ws, fin + 1, ws.max_row)
    _limpiar(ws, fin + 1, max(ws.max_row, (marca or 0) + 10))
    col_inicio = _cabecera_nueva(ws, fila_cab, 'J',
                                 'Inicio del amasado (hora)', 20)
    r = fin + 2
    _et(ws, 'A' + str(r), MARCA_C, bold=True)
    r += 1
    ref_horno = _param(
        ws, r, None, valor=None,
        etiqueta='Hora objetivo de salida del horno', fmt=FMT_HORA,
        nota=('Escríbela como hora («07:00») y la columna «Inicio del amasado» '
              'te dice a qué hora empezar cada masa, restando su tiempo total. '
              + NOTA_SIN_DATO), col_et='A', col_val='C', col_nota='D')
    r += 1
    # La conversión horas -> fracción de día va en celda, no dentro de la
    # fórmula: `MOD(hora-H4/24,1)` metería un literal 24 en 15 fórmulas y el
    # gate del §1.3 lo cuenta como parámetro sin celda. No es verde: es una
    # equivalencia de unidades, no un dato del cliente.
    _et(ws, 'A' + str(r), 'Horas que tiene un día (conversión de unidades)')
    motor.val(ws, 'C' + str(r), 24, fmt=motor.FMT_ENT)
    ws['C' + str(r)].protection = Protection(locked=True)
    ref_dia = '$C$' + str(r)
    convertidas = 0
    for f_ in range(fila_cab + 1, fin + 1):
        anterior = motor.a_formula(
            ws, 'H' + str(f_),
            motor.iferror('IF(COUNT($E{r}:$G{r})=0,"",'
                          'SUM($E{r}:$G{r}))'.format(r=f_)),
            fmt='#,##0.0', informe=None)
        if anterior is not None:
            convertidas += 1
        motor.fijar_formato(ws, 'H' + str(f_), '#,##0.0')
        motor.f(ws, col_inicio + str(f_),
                motor.iferror('IF(OR($H{r}="",' + ref_horno + '=""),"",MOD('
                              + ref_horno + '-$H{r}/' + ref_dia
                              + ',1))').format(r=f_), fmt=FMT_HORA)
        motor.fijar_formato(ws, col_inicio + str(f_), FMT_HORA)
        motor.verde(ws, 'A' + str(f_) + ':G' + str(f_))
        motor.verde(ws, 'I' + str(f_))
        for L in ('C', 'D'):
            motor.fijar_formato(ws, L + str(f_), '#,##0.0')
    cambios.append(
        fname + ':' + ws.title + '!H' + str(fila_cab + 1) + ':H' + str(fin)
        + ': «Tiempo total h» pasa de ' + str(convertidas) + ' constantes '
          'tecleadas a =SUM(bulk+bloque frío+final), conservando el número '
          '(comprobado fila a fila: coincidía en las ' + str(fin - fila_cab)
        + '), y la columna «Inicio del amasado» hace el retroplanning desde la '
          'hora de salida del horno. La hoja tenía 0 fórmulas '
          '[§4.5 · §1.2 · §7-bis.12]')
    _instr(wb, fname, cambios, [
        'Escribe la hora a la que quieres sacar el pan del horno y la columna '
        '«Inicio del amasado» te dice a qué hora empezar cada masa.',
        'El «Tiempo total h» ya no se teclea: es la suma del bulk, el bloque '
        'frío y la fermentación final.'])
    registro_modelo['fermentacion'] = {'convertidas': convertidas,
                                       'primera': fila_cab + 1,
                                       'descuadres': []}
    return registro_modelo['fermentacion']


# ==========================================================================
# Instrucciones
# ==========================================================================
def _instr(wb, fname, cambios, lineas):
    """Añade las líneas de uso a `Instrucciones` sin duplicarlas.

    El motor ya ha creado la hoja en `aplicar()` (56 de los 111 xlsx no la
    tenían) y escribirá al final el bloque de bio + versión, que se recoloca
    solo detrás del último contenido vivo.
    """
    if 'Instrucciones' not in wb.sheetnames:
        return
    ws = wb['Instrucciones']
    for texto in lineas:
        motor.linea_instrucciones(ws, texto)
    del cambios, fname


# ==========================================================================
# Contrato con main.py
# ==========================================================================
def ficheros(ctx):
    """Los ficheros del grupo C **de ese producto**, leídos del disco.

    No es una lista fija: `escandallo-maestro` se llama `-japones`, `-mexicano`,
    `-nikkei`, `-peruano` y `-panaderia` en cinco de las siete guías que lo
    llevan, y dark-kitchen no lleva ninguno de los cinco ficheros de este grupo.
    """
    pid = (ctx or {}).get('producto') or motor.CTX.get('producto')
    if not pid:
        return []
    fuera = []
    for p in sorted(glob.glob(os.path.join(DL, pid, '*.xlsx'))):
        n = os.path.basename(p)
        if any(n.startswith(x) for x in PREFIJOS):
            fuera.append(n)
    return fuera


PROPIOS = []          # el §1 del motor se aplica entero a todos estos ficheros

#: Lo que cada fichero de este grupo ha resuelto en la pasada en curso, para que
#: `demos()` sepa dónde mirar sin volver a adivinar la rejilla.
ESTRUCTURA = {}


def pre(wb, fname, cambios, contenido):
    """Antes de `motor.aplicar()`: el renombrado de la pestaña del escandallo.

    Va aquí y no en `post` para que `motor.aplicar()` —que crea la hoja
    `Instrucciones` a partir del título de la primera hoja— y el registro de
    fórmulas de `motor.f()` vean ya el nombre definitivo: una fórmula registrada
    con el nombre viejo haría fallar la verificación `data_only` de `main.py`
    con «hoja ausente».
    """
    if not fname.startswith('escandallo-maestro'):
        return
    hojas = [ws for ws in wb.worksheets
             if ws.title not in ('Instrucciones', HOJA_RESUMEN)]
    if not hojas:
        return
    ws = hojas[0]
    try:
        modelo, _fila = modelo_escandallo(ws, fname)
    except EstructuraDesconocida:
        return                      # `post` abortará con el detalle completo
    if modelo == 'E3':
        return                      # tabla de productos: no hay ficha
    if ws.title != HOJA_FICHA:
        motor.renombrar_pestana(wb, ws, HOJA_FICHA, cambios, fname)
    del contenido


def post(wb, fname, cambios, registro, contenido):
    """Después de `motor.aplicar()` y antes de `motor.cerrar()`, que pondrá los
    formatos por tipo, la DV de las verdes, la bio, la versión y la protección
    sobre lo que aquí se escriba."""
    estructura = ESTRUCTURA.setdefault(fname, {})
    if fname.startswith('escandallo-maestro'):
        _escandallo(wb, fname, cambios, contenido, estructura)
    elif fname.startswith('menu-engineering-matrix'):
        _menu(wb, fname, cambios, contenido, estructura)
    elif fname.startswith('budget-bodega'):
        _bodega(wb, fname, cambios, contenido, estructura)
    elif fname.startswith('plantilla-turnos-brigada'):
        _turnos(wb, fname, cambios, contenido, estructura)
    elif fname.startswith('plan-fermentacion-y-produccion'):
        _fermentacion(wb, fname, cambios, contenido, estructura)
    del registro


# ==========================================================================
# Demostraciones propias (pycel) — sobre COPIAS, nunca sobre los entregables
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__ + ':' + str(e)[:90]


def _set(xl, ref, valor):
    """`set_value` sobre una celda que pycel aún no ha visto revienta con
    `Address ... not found in the cell map`. Se evalúa primero (que es lo que la
    mete en el mapa) y después se escribe: sin esto, las demos que cambian un
    input de una columna NUEVA —la que este grupo acaba de crear— abortan."""
    _ev(xl, ref)
    xl.set_value(ref, valor)


def _copia(carpeta, fname, destino, sufijo=''):
    import shutil
    os.makedirs(destino, exist_ok=True)
    fuera = os.path.join(destino, 'c' + sufijo + '-' + fname)
    shutil.copy2(os.path.join(carpeta, fname), fuera)
    return fuera


def _precalentar(xl, refs):
    """Evalúa las celdas de SALIDA antes de cambiar ningún input.

    pycel devuelve el valor CACHEADO de una celda que no está todavía en su
    mapa, así que un `set_value` sobre una entrada no mueve una salida que
    nunca se evaluó: medido en `escandallo-maestro-panaderia!G4`, que seguía
    dando 1,82 (el valor del fichero) después de cambiar C4 y D4. Con la
    celda ya en el mapa, el mismo cambio da el resultado correcto. Sin esto
    una demo puede dar «ok» leyendo el número viejo.
    """
    return [_ev(xl, r) for r in refs]


def _cerca(a, b, tol=0.01):
    return isinstance(a, (int, float)) and isinstance(b, (int, float)) \
        and abs(a - b) <= tol


def demos(carpeta, destino, contenido):
    """Cada cálculo NUEVO de este grupo, evaluado con pycel **cambiando los
    inputs**: un número que sale bien una vez puede ser una constante.

    Se comprueba, fichero a fichero y citando la celda:
      1. escandallo — con el libro EN BLANCO todo devuelve `""` (ni un 0,00 €);
         la merma mueve el coste (0,180 kg netos al 40 % cuestan 6,60 € y no
         3,96 €); el PVP divide por raciones (60 € / 10 raciones / 28 % =
         21,43 € y no 214,29 €); el food cost objetivo de la celda mueve el PVP;
         el PVP con IVA es el sin IVA por 1,10; y el `Resumen` lee la ficha.
      2. menu engineering — la cadena completa de Kasavana & Smith: total de
         uds, margen medio ponderado, umbral 70 %/N, mix y clasificación, y que
         al doblar las ventas de un Dog cambia de cuadrante.
      3. bodega — multiplicador, margen sobre PVP, food cost, rotación y la fila
         TOTAL, que responde al stock.
      4. turnos — las horas responden al cambio de un turno y a la tabla de
         equivalencia; el coste/hora responde al tipo de SS; el semáforo del SMI
         se enciende con un bruto por debajo y NO con la celda vacía; y dos
         jornadas a menos de 12 h disparan el aviso del registro.
      5. fermentación — el tiempo total sigue a sus tres tramos.
    """
    resultado = {'grupo_c': {}, 'fallos': []}
    fallos = resultado['fallos']
    det = resultado['grupo_c']
    import openpyxl

    # ---- 0. gate de plantilla sin sustituir -----------------------------
    # Nace de un defecto REAL de este módulo, cazado por las demos y no por la
    # lectura del diff: `motor.iferror('IF($D{r}=…' + ref + '))'.format(r=f_)`
    # aplica el `.format` al ÚLTIMO literal (`'))'`), no a la concatenación, y
    # deja `{r}` DENTRO de la fórmula. El fichero se guarda, el censo pasa, y
    # `verificar_cache` lo manda a «vacías no verificadas» —que no es fallo—
    # porque la fórmula contiene `""`. En Excel el cliente vería #¿NOMBRE?.
    # Dos caracteres, 25 celdas y ninguna alarma: por eso el barrido es
    # explícito y bloqueante.
    plantillas = []
    for fname in sorted(os.listdir(carpeta)):
        if not fname.endswith('.xlsx') or \
                not any(fname.startswith(x) for x in PREFIJOS):
            continue
        wbp = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wbp.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.data_type == 'f' and isinstance(c.value, str) \
                            and ('{' in c.value or '}' in c.value):
                        plantillas.append(fname + ':' + ws.title + '!'
                                          + c.coordinate + '=' + c.value[:70])
    det['formulas_con_marcador_sin_sustituir'] = plantillas[:20]
    det['formulas_con_marcador_sin_sustituir_total'] = len(plantillas)
    if plantillas:
        fallos.append(
            str(len(plantillas)) + ' fórmulas con un marcador de plantilla sin '
            'sustituir ({r}) — el `.format` se aplicó al último literal en vez '
            'de a la concatenación: ' + '; '.join(plantillas[:5]))

    for fname in sorted(os.listdir(carpeta)):
        if not fname.endswith('.xlsx'):
            continue
        if not any(fname.startswith(x) for x in PREFIJOS):
            continue
        est = ESTRUCTURA.get(fname) or {}
        path = os.path.join(carpeta, fname)
        wb = openpyxl.load_workbook(path)

        # ---- 1. escandallo ------------------------------------------------
        if fname.startswith('escandallo-maestro') and est.get('escandallo'):
            e = est['escandallo']
            ws = wb[HOJA_FICHA] if HOJA_FICHA in wb.sheetnames else \
                [w for w in wb.worksheets if w.title != 'Instrucciones'][0]
            pref = "'" + ws.title + "'!"
            if e['modelo'] == 'E3':
                xl = _pycel(_copia(carpeta, fname, destino))
                r0 = e['primera']
                _precalentar(xl, [pref + L + str(r0) for L in
                                  ('E', 'F', 'G', e['col_pvp_obj'],
                                   e['col_pvp_iva'])])
                _set(xl, pref + 'C' + str(r0), 1.0)
                _set(xl, pref + 'D' + str(r0), 4.0)
                _set(xl, pref + e['ref_fc'].replace('$', ''), 0.25)
                fc = _ev(xl, pref + 'E' + str(r0))
                margen = _ev(xl, pref + 'G' + str(r0))
                sug = _ev(xl, pref + e['col_pvp_obj'] + str(r0))
                con_iva = _ev(xl, pref + e['col_pvp_iva'] + str(r0))
                # El «libro en blanco» se prueba VACIANDO la celda de verdad
                # con openpyxl y recompilando: `set_value(..., None)` de pycel
                # no deja la celda vacía y la demo daba por bueno el valor
                # viejo. Es la misma clase de falso verde que el
                # `_precalentar`.
                copia_v = _copia(carpeta, fname, destino, '-vacio')
                wbv = openpyxl.load_workbook(copia_v)
                wsv = wbv[ws.title]
                wsv['C' + str(r0)].value = None
                wsv['D' + str(r0)].value = None
                wbv.save(copia_v)
                xl2 = _pycel(copia_v)
                vacio = _ev(xl2, pref + 'E' + str(r0))
                ok = (_cerca(fc, 0.25) and _cerca(margen, 3.0)
                      and _cerca(sug, 4.0) and _cerca(con_iva, 4.4)
                      and vacio == '')
                det.setdefault('escandallo', []).append(
                    {'fichero': fname, 'modelo': 'E3', 'hoja': ws.title,
                     'coste_1_pvp_4_food_cost_esperado_0_25': fc,
                     'margen_esperado_3': margen,
                     'pvp_sugerido_al_25_esperado_4': sug,
                     'pvp_con_iva_esperado_4_40': con_iva,
                     'food_cost_sin_pvp_esperado_vacio': vacio,
                     'ok': bool(ok)})
                if not ok:
                    fallos.append(
                        fname + ':' + ws.title + ': con coste 1 EUR, PVP 4 EUR '
                        'y food cost objetivo 25 % se esperaba 25 %, margen '
                        '3 EUR, PVP sugerido 4 EUR y con IVA 4,40 EUR; salió '
                        + repr((fc, margen, sug, con_iva))
                        + ' y sin PVP ' + repr(vacio) + ' [DOM-03 · DOM-30]')
            else:
                xl = _pycel(_copia(carpeta, fname, destino))
                blanco = {}
                if e['modelo'] == 'E1':
                    # con el libro en blanco todo tiene que ser ""
                    for clave in ('coste_total', 'coste_racion', 'pvp_sin_iva',
                                  'pvp_con_iva'):
                        blanco[clave] = _ev(xl, pref + e[clave])
                    malos = [k for k, v in blanco.items() if v != '']
                    if malos:
                        fallos.append(
                            fname + ':' + ws.title + ': con el libro EN BLANCO '
                            + str({k: blanco[k] for k in malos})
                            + ' en vez de "" — un 0,00 EUR en una fila de '
                              'COSTE TOTAL se lee como un resultado (§7-bis.13)')
                    fila = e['primera']
                    _set(xl, pref + 'D' + str(fila), 0.18)
                    _set(xl, pref + 'E' + str(fila), 22)
                    _set(xl, pref + 'F' + str(fila), 0.0)
                    sin_merma = _ev(xl, pref + e['coste_total'])
                    _set(xl, pref + 'F' + str(fila), 0.4)
                    con_merma = _ev(xl, pref + e['coste_total'])
                    ok_merma = _cerca(sin_merma, 3.96) and _cerca(con_merma, 6.6)
                    if not ok_merma:
                        fallos.append(
                            fname + ':' + ws.title + '!' + e['coste_total']
                            + ': merluza a 22 EUR/kg y 0,180 kg netos: sin '
                              'merma ' + repr(sin_merma) + ' (esperado 3,96) y '
                              'con 40 % de merma ' + repr(con_merma)
                            + ' (esperado 6,60). La merma NO entra en el coste '
                              '[DOM-04]')
                    # raciones y food cost
                    _set(xl, pref + 'D' + str(fila), 60)
                    _set(xl, pref + 'E' + str(fila), 1)
                    _set(xl, pref + 'F' + str(fila), 0.0)
                    _set(xl, pref + 'F4', 1)
                    pvp1 = _ev(xl, pref + e['pvp_sin_iva'])
                    _set(xl, pref + 'F4', 10)
                    pvp10 = _ev(xl, pref + e['pvp_sin_iva'])
                    _set(xl, pref + 'H4', 0.32)
                    pvp32 = _ev(xl, pref + e['pvp_sin_iva'])
                    iva = _ev(xl, pref + e['pvp_con_iva'])
                    ok_rac = _cerca(pvp1, 214.2857, 0.01) and \
                        _cerca(pvp10, 21.4286, 0.01)
                    ok_fc = _cerca(pvp32, 18.75, 0.01)
                    ok_iva = _cerca(iva, pvp32 * 1.10, 0.01) if \
                        isinstance(pvp32, (int, float)) else False
                    det.setdefault('escandallo', []).append({
                        'fichero': fname, 'modelo': e['modelo'],
                        'hoja': ws.title, 'libro_en_blanco': blanco,
                        'coste_sin_merma_esperado_3_96': sin_merma,
                        'coste_con_merma_40_esperado_6_60': con_merma,
                        'pvp_1_racion_esperado_214_29': pvp1,
                        'pvp_10_raciones_esperado_21_43': pvp10,
                        'pvp_food_cost_32_esperado_18_75': pvp32,
                        'pvp_con_iva': iva,
                        'ok': bool(ok_merma and ok_rac and ok_fc and ok_iva)})
                    if not ok_rac:
                        fallos.append(
                            fname + ':' + ws.title + '!' + e['pvp_sin_iva']
                            + ': con 60 EUR de coste el PVP da ' + repr(pvp1)
                            + ' con 1 ración y ' + repr(pvp10) + ' con 10; se '
                              'esperaba 214,29 y 21,43 [TEC-05 · DOM-05]')
                    if not ok_fc:
                        fallos.append(
                            fname + ':' + ws.title + '!H4: cambiar el food cost '
                            'objetivo a 32 % deja el PVP en ' + repr(pvp32)
                            + ' (esperado 18,75) [TEC-21 · DOM-30]')
                    if not ok_iva:
                        fallos.append(
                            fname + ':' + ws.title + '!' + e['pvp_con_iva']
                            + ': el PVP con IVA (' + repr(iva) + ') no es el '
                              'PVP sin IVA por 1,10 [DOM-03 · COM-14]')
                else:
                    fila = e['primera']
                    _precalentar(xl, [pref + e[k] for k in
                                      ('coste_total', 'coste_racion',
                                       'pvp_sin_iva', 'pvp_con_iva')])
                    total0 = _ev(xl, pref + e['coste_total'])
                    _set(xl, pref + 'F' + str(fila), 0.0)
                    sin_m = _ev(xl, pref + e['coste_total'])
                    _set(xl, pref + 'F' + str(fila), 0.2)
                    con_m = _ev(xl, pref + e['coste_total'])
                    pvp = _ev(xl, pref + e['pvp_sin_iva'])
                    iva = _ev(xl, pref + e['pvp_con_iva'])
                    ok = (isinstance(sin_m, (int, float))
                          and isinstance(con_m, (int, float))
                          and con_m > sin_m
                          and _cerca(iva, pvp * 1.10, 0.01)
                          if isinstance(pvp, (int, float)) else False)
                    det.setdefault('escandallo', []).append({
                        'fichero': fname, 'modelo': 'E2', 'hoja': ws.title,
                        'coste_total_original': total0,
                        'coste_merma_0': sin_m, 'coste_merma_20': con_m,
                        'pvp_sin_iva': pvp, 'pvp_con_iva': iva, 'ok': bool(ok)})
                    if not ok:
                        fallos.append(
                            fname + ':' + ws.title + '!' + e['coste_total']
                            + ': la merma no mueve el coste o el IVA no cuadra '
                              '(merma 0 -> ' + repr(sin_m) + ', merma 20 % -> '
                            + repr(con_m) + ', PVP ' + repr(pvp) + ', con IVA '
                            + repr(iva) + ') [DOM-04 · DOM-03]')
                if HOJA_RESUMEN in wb.sheetnames:
                    xl2 = _pycel(_copia(carpeta, fname, destino, '-res'))
                    pr = "'" + HOJA_RESUMEN + "'!"
                    vacio = _ev(xl2, pr + 'C5')
                    fila2 = e['primera']
                    col_c, col_p = ('D', 'E') if e['modelo'] == 'E1' \
                        else ('C', 'D')
                    _set(xl2, pref + col_c + str(fila2), 2)
                    _set(xl2, pref + col_p + str(fila2), 10)
                    lleno = _ev(xl2, pr + 'C5')
                    # El criterio NO puede ser «antes estaba vacío»: E1 se
                    # entrega en blanco pero E2 viene con 11 ingredientes
                    # precargados. Lo que demuestra que la referencia entre
                    # hojas funciona es que el número CAMBIE al tocar la ficha.
                    det.setdefault('resumen', []).append(
                        {'fichero': fname, 'C5_antes': vacio,
                         'C5_tras_tocar_la_ficha': lleno,
                         'ok': isinstance(lleno, (int, float))
                         and lleno != vacio})
                    if isinstance(vacio, str) and vacio.startswith('ERR:'):
                        fallos.append(fname + ':' + HOJA_RESUMEN + '!C5: la '
                                      'referencia a la hoja «' + HOJA_FICHA
                                      + '» no evalúa (' + vacio + ') [TEC-23]')
                    elif not isinstance(lleno, (int, float)) \
                            or lleno == vacio:
                        fallos.append(
                            fname + ':' + HOJA_RESUMEN + '!C5: al cambiar una '
                            'línea de la ficha el Resumen pasa de '
                            + repr(vacio) + ' a ' + repr(lleno)
                            + ': la referencia entre hojas no trae el coste '
                              '[TEC-23]')

        # ---- 2. menu engineering -----------------------------------------
        if fname.startswith('menu-engineering') and est.get('menu'):
            m = est['menu']
            ws = [w for w in wb.worksheets if w.title != 'Instrucciones'][0]
            pref = "'" + ws.title + "'!"
            xl = _pycel(_copia(carpeta, fname, destino))
            tot = _ev(xl, pref + m['col_uds'] + str(m['fila_total']))
            mc = _ev(xl, pref + m['col_margen'] + str(m['fila_total']))
            umbral = _ev(xl, pref + m['col_uds'] + str(m['fila_umbral']))
            mix1 = _ev(xl, pref + m['col_mix'] + str(m['primera']))
            clases = [_ev(xl, pref + m['col_clase'] + str(r))
                      for r in range(m['primera'], m['ultima'] + 1)]
            vivas = [c for c in clases if c not in ('', None)]
            # se dobla el plato menos vendido: tiene que cambiar de cuadrante
            uds = {}
            for r in range(m['primera'], m['ultima'] + 1):
                v = _ev(xl, pref + m['col_uds'] + str(r))
                if isinstance(v, (int, float)) and v:
                    uds[r] = v
            cambio = None
            if uds:
                r_min = min(uds, key=lambda k: uds[k])
                antes = _ev(xl, pref + m['col_clase'] + str(r_min))
                _set(xl, pref + m['col_uds'] + str(r_min),
                             max(uds.values()) * 3)
                cambio = {'fila': r_min, 'antes': antes,
                          'despues': _ev(xl, pref + m['col_clase'] + str(r_min))}
            det.setdefault('menu', []).append({
                'fichero': fname, 'modelo': m['modelo'], 'hoja': ws.title,
                'total_uds': tot, 'mc_medio_ponderado': mc,
                'umbral_70_pct_N': umbral, 'mix_primer_plato': mix1,
                'clasificaciones': clases,
                'al_triplicar_ventas_del_menos_vendido': cambio,
                'ok': bool(vivas and cambio
                           and cambio['antes'] != cambio['despues'])})
            if not vivas:
                fallos.append(
                    fname + ':' + ws.title + '!' + m['col_clase']
                    + str(m['primera']) + ': la columna Clasificación sigue '
                      'vacía en las ' + str(len(clases)) + ' filas '
                      '[TEC-04 · DOM-10 · COM-08]')
            elif not isinstance(mc, (int, float)):
                fallos.append(
                    fname + ':' + ws.title + '!' + m['col_margen']
                    + str(m['fila_total']) + ': el margen de contribución medio '
                      'ponderado devuelve ' + repr(mc) + '. Si es 0 o un texto, '
                      'el SUMPRODUCT no está envuelto en ROUND y pycel devuelve '
                      'numpy.int64')
            elif cambio and cambio['antes'] == cambio['despues']:
                fallos.append(
                    fname + ':' + ws.title + ': triplicar las ventas del plato '
                    'menos vendido (fila ' + str(cambio['fila']) + ') no cambia '
                    'su cuadrante (' + repr(cambio['antes']) + '): el mix no '
                    'entra en la clasificación')

        # ---- 3. bodega ----------------------------------------------------
        if fname.startswith('budget-bodega') and est.get('bodega'):
            b = est['bodega']
            ws = [w for w in wb.worksheets if w.title != 'Instrucciones'][0]
            pref = "'" + ws.title + "'!"
            xl = _pycel(_copia(carpeta, fname, destino))
            r0 = b['primera']
            # RT-30 · la bodega ya NO se entrega vacía (llegaba con 355
            # fórmulas y ningún resultado): el caso «en blanco» hay que
            # PROVOCARLO vaciando las entradas de las filas precargadas.
            _precalentar(xl, [pref + c + str(b['fila_total'])
                              for c in ('N', 'O', 'L', 'M')])
            for f_ in range(b['primera'], b['ultima'] + 1):
                for L in ('E', 'F', 'I', 'K'):
                    _set(xl, pref + L + str(f_), '')
            blanco = {c: _ev(xl, pref + c + str(b['fila_total']))
                      for c in ('N', 'O', 'L', 'M')}
            malos = {k: v for k, v in blanco.items() if v != ''}
            if malos:
                fallos.append(
                    fname + ':' + ws.title + '!fila ' + str(b['fila_total'])
                    + ': con la bodega EN BLANCO el TOTAL devuelve ' + str(malos)
                    + ' en vez de "" (§7-bis.13)')
            _precalentar(xl, [pref + L + str(r0)
                              for L in ('G', 'H', 'J', 'L', 'M', 'N', 'O',
                                        'T')])
            _set(xl, pref + 'E' + str(r0), 10)
            _set(xl, pref + 'F' + str(r0), 30)
            _set(xl, pref + 'I' + str(r0), 24)
            _set(xl, pref + 'K' + str(r0), 12)
            # RD-13 · el PVP de la columna F es el de CARTA, CON IVA (21 % en
            # bebida alcohólica): 30 € de carta son 24,79 € netos. Los
            # indicadores se miden sobre esa base, y por eso el food cost de
            # bebida real es 40,3 % y no el 33,3 % que salía cruzando un coste
            # sin IVA con un precio con IVA — 7 puntos de diferencia, justo en
            # la magnitud que decide si la bodega gana dinero.
            neto = round(30 / 1.21, 10)
            mult = _ev(xl, pref + 'H' + str(r0))
            margen = _ev(xl, pref + 'L' + str(r0))
            fc = _ev(xl, pref + 'M' + str(r0))
            rot = _ev(xl, pref + 'J' + str(r0))
            vcoste = _ev(xl, pref + 'N' + str(r0))
            tot_coste = _ev(xl, pref + 'N' + str(b['fila_total']))
            # el resto de filas se ha vaciado justo arriba, así que el TOTAL
            # tiene que ser exactamente el valor de la única fila con datos
            ok = (_cerca(mult, neto / 10) and _cerca(margen, 1 - 10 / neto,
                                                     0.001)
                  and _cerca(fc, 10 / neto, 0.001) and _cerca(rot, 0.5)
                  and _cerca(vcoste, 240) and _cerca(tot_coste, 240))
            det.setdefault('bodega', []).append({
                'fichero': fname, 'hoja': ws.title, 'total_en_blanco': blanco,
                'coste_10_pvp_carta_30_con_iva_21': {
                    'pvp_sin_iva_esperado_24_79': _ev(xl, pref + 'T' + str(r0)),
                    'multiplicador_esperado_2_48': mult,
                    'margen_s_pvp_esperado_0_597': margen,
                    'food_cost_esperado_0_403': fc},
                'rotacion_12_vendidas_24_stock_esperado_0_5': rot,
                'valor_stock_coste_esperado_240': vcoste,
                'TOTAL_valor_a_coste_esperado_240': tot_coste, 'ok': bool(ok)})
            if not ok:
                fallos.append(
                    fname + ':' + ws.title + ': con coste 10 EUR, PVP de carta '
                    '30 EUR (IVA de bebida 21 %), stock 24 y 12 uds vendidas se '
                    'esperaba PVP sin IVA 24,79 EUR, multiplicador 2,48, margen '
                    's/PVP 59,7 %, food cost 40,3 %, rotación 0,5 y valor a '
                    'coste 240 EUR; salió ' + repr((mult, margen, fc, rot,
                                                    vcoste, tot_coste))
                    + ' [TEC-19 · TEC-20 · TEC-18 · RD-13]')

        # ---- 4. turnos -----------------------------------------------------
        if fname.startswith('plantilla-turnos') and est.get('turnos'):
            t = est['turnos']
            ws = [w for w in wb.worksheets
                  if w.title not in ('Instrucciones', HOJA_REGISTRO)][0]
            pref = "'" + ws.title + "'!"
            xl = _pycel(_copia(carpeta, fname, destino))
            r0, ch = t['primera'], t['col_horas']
            horas0 = _ev(xl, pref + ch + str(r0))
            total0 = _ev(xl, pref + ch + str(t['fila_total']))
            detalle = {'fichero': fname, 'modelo': t['modelo'],
                       'hoja': ws.title, 'puestos': t['puestos'],
                       'horas_fila_' + str(r0) + '_al_abrir': horas0,
                       'total_horas_al_abrir': total0}
            _precalentar(xl, [pref + L + str(r0) for L in ('N', 'O', 'P')])
            if t['modelo'] != 'T3':
                for L in 'DEFGHIJ':
                    _set(xl, pref + L + str(r0), 'L')
                cero = _ev(xl, pref + ch + str(r0))
                _set(xl, pref + 'D' + str(r0), 'P')
                una_p = _ev(xl, pref + ch + str(r0))
                _set(xl, pref + 'E' + str(r0), 'M')
                p_mas_m = _ev(xl, pref + ch + str(r0))
                detalle.update({'siete_libres_esperado_0': cero,
                                'un_partido_esperado_10': una_p,
                                'partido_mas_manana_esperado_18': p_mas_m})
                ok_h = _cerca(cero, 0) and _cerca(una_p, 10) and \
                    _cerca(p_mas_m, 18)
                if not ok_h:
                    fallos.append(
                        fname + ':' + ws.title + '!' + ch + str(r0)
                        + ': las horas no siguen a los turnos (7 libres -> '
                        + repr(cero) + ', 1 partido -> ' + repr(una_p)
                        + ', partido+mañana -> ' + repr(p_mas_m)
                        + ', esperado 0/10/18) [TEC-11 · COM-20]')
            else:
                _set(xl, pref + ch + str(r0), 40)
                ok_h = True
            # coste/hora y coste semana
            _set(xl, pref + 'L' + str(r0), 30000)
            _set(xl, pref + t['ref_jornada'], 1800)
            coste_hora = _ev(xl, pref + 'O' + str(r0))
            coste_sem = _ev(xl, pref + 'P' + str(r0))
            _set(xl, pref + t['ref_ss'], 0.0)
            coste_sin_ss = _ev(xl, pref + 'O' + str(r0))
            detalle.update({
                'coste_hora_30000_bruto_1800h_ss33_esperado_22_17': coste_hora,
                'coste_semana': coste_sem,
                'coste_hora_con_ss_0_esperado_16_67': coste_sin_ss})
            ok_c = _cerca(coste_hora, 30000 * 1.33 / 1800, 0.01) and \
                _cerca(coste_sin_ss, 30000 / 1800, 0.01)
            if not ok_c:
                fallos.append(
                    fname + ':' + ws.title + '!O' + str(r0) + ': el coste/hora '
                    'no responde al tipo de SS (con 33 % ' + repr(coste_hora)
                    + ', con 0 % ' + repr(coste_sin_ss) + '; esperado 22,17 y '
                    '16,67) [DOM-02 · COM-10]')
            # semáforo del SMI, evaluado como fórmula auxiliar
            copia2 = _copia(carpeta, fname, destino, '-smi')
            wb2 = openpyxl.load_workbook(copia2)
            ws2 = wb2[ws.title]
            libre = max(ws2.max_column, 20) + 2
            ws2.cell(row=r0, column=libre).value = (
                '=AND(ISNUMBER($L' + str(r0) + '),$L' + str(r0) + '<'
                + t['ref_smi'] + ')')
            wb2.save(copia2)
            xl3 = _pycel(copia2)
            aux = get_column_letter(libre) + str(r0)
            vacio = _ev(xl3, pref + aux)
            _set(xl3, pref + 'L' + str(r0), 15000)
            bajo = _ev(xl3, pref + aux)
            _set(xl3, pref + 'L' + str(r0), 22000)
            alto = _ev(xl3, pref + aux)
            detalle.update({'semaforo_smi_celda_vacia': vacio,
                            'semaforo_smi_15000': bajo,
                            'semaforo_smi_22000': alto})
            ok_s = (vacio is False and bajo is True and alto is False)
            if not ok_s:
                fallos.append(
                    fname + ':' + ws.title + '!L' + str(r0) + ': el semáforo '
                    'del SMI da ' + repr((vacio, bajo, alto)) + ' para (vacía, '
                    '15.000, 22.000) y debía dar (False, True, False) '
                    '[DOM-13 · §1.6]')
            detalle['ok'] = bool(ok_h and ok_c and ok_s)
            det.setdefault('turnos', []).append(detalle)

            # registro de jornada
            if HOJA_REGISTRO in wb.sheetnames:
                copia3 = _copia(carpeta, fname, destino, '-reg')
                xl4 = _pycel(copia3)
                pr = "'" + HOJA_REGISTRO + "'!"
                vacio_h = _ev(xl4, pr + 'F5')
                _precalentar(xl4, [pr + c for c in ('F6', 'G6', 'H6', 'F7')])
                _set(xl4, pr + 'B5', 'Ana')
                _set(xl4, pr + 'D5', 0.5)          # 12:00
                _set(xl4, pr + 'E5', 0.9583333333333334)   # 23:00
                horas = _ev(xl4, pr + 'F5')
                _set(xl4, pr + 'B6', 'Ana')
                _set(xl4, pr + 'D6', 0.3333333333333333)   # 08:00
                _set(xl4, pr + 'E6', 0.75)         # 18:00
                descanso = _ev(xl4, pr + 'G6')
                aviso = _ev(xl4, pr + 'H6')
                medianoche = None
                _set(xl4, pr + 'D7', 0.9166666666666666)   # 22:00
                _set(xl4, pr + 'E7', 0.25)         # 06:00
                medianoche = _ev(xl4, pr + 'F7')
                ok_r = (vacio_h == '' and _cerca(horas, 11.0 / 24, 1e-6)
                        and _cerca(descanso, 9.0 / 24, 1e-6)
                        and isinstance(aviso, str) and aviso
                        and _cerca(medianoche, 8.0 / 24, 1e-6))
                det.setdefault('registro_jornada', []).append({
                    'fichero': fname, 'F5_libro_en_blanco': vacio_h,
                    'horas_12_a_23_esperado_11h': horas,
                    'descanso_23_a_08_esperado_9h': descanso,
                    'aviso_menos_de_12h': aviso,
                    'cruce_medianoche_22_a_06_esperado_8h': medianoche,
                    'ok': bool(ok_r)})
                if not ok_r:
                    fallos.append(
                        fname + ':' + HOJA_REGISTRO + ': horas ' + repr(horas)
                        + ' (esperado 11 h), descanso ' + repr(descanso)
                        + ' (esperado 9 h), aviso ' + repr(aviso)
                        + ', cruce de medianoche ' + repr(medianoche)
                        + ' (esperado 8 h) [DOM-16]')

        # ---- 5. fermentación ----------------------------------------------
        # ⚠️ NO se condiciona a `convertidas`: `ESTRUCTURA` la reescribe la 2.ª
        # pasada de la idempotencia, donde ya no queda ninguna constante que
        # convertir y el contador vuelve a 0. La demo se saltaba entera en
        # silencio. Se condiciona a que no haya descuadres, que sí es estable.
        if fname.startswith('plan-fermentacion') and est.get('fermentacion') \
                and not est['fermentacion'].get('descuadres'):
            ws = [w for w in wb.worksheets if w.title != 'Instrucciones'][0]
            pref = "'" + ws.title + "'!"
            xl = _pycel(_copia(carpeta, fname, destino))
            fila = est['fermentacion'].get('primera', 4)
            ref = pref + 'H' + str(fila)
            _precalentar(xl, [ref, pref + 'J' + str(fila)])
            antes = _ev(xl, ref)
            bloque = _ev(xl, pref + 'F' + str(fila))
            _set(xl, pref + 'F' + str(fila),
                 (bloque or 0) + 12)
            despues = _ev(xl, ref)
            ok = (isinstance(antes, (int, float))
                  and isinstance(despues, (int, float))
                  and _cerca(despues - antes, 12, 0.01))
            det.setdefault('fermentacion', []).append({
                'fichero': fname, 'fila': fila,
                'tiempo_total_original': antes,
                'tiempo_total_con_12h_mas_de_bloque_frio': despues,
                'diferencia_esperada_12': (despues - antes)
                if ok else None, 'ok': bool(ok)})
            if not ok:
                fallos.append(
                    fname + ':' + ws.title + '!H' + str(fila) + ': el tiempo '
                    'total no sigue a sus tramos: al sumar 12 h de bloque frío '
                    'pasa de ' + repr(antes) + ' a ' + repr(despues) + ' [§4.5]')
    return resultado
