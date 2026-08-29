#!/usr/bin/env python3
"""
grupo_a.py — §2 de `guias-v2-SPEC.md`: las **plantillas financieras** de la
familia «Guías Cómo Montar» v2.0.

Ficheros (§2): `calculadora-ticket-medio.xlsx`, `pl-mensual-escenarios.xlsx`,
`plan-financiero-3-anos.xlsx`, `cash-flow-break-even.xlsx` y
`calculadora-capex.xlsx`, en las **7** guías que los llevan.
`guia-dark-kitchen` NO entra: su `calculadora-viabilidad-dark-kitchen.xlsx`
ya calcula y **sólo recibe §1** (SPEC §2, cabecera) — por eso `ficheros()` es
una función y devuelve `[]` para ese pid.

REGLA DE FAMILIA (la razón de ser de este módulo): **nada por posición fija sin
comprobar la etiqueta**. Medido el 2026-08-29 abriendo los ficheros vivos:

  · `pl-mensual-escenarios!'Pesimista'!EBITDA` está en la fila **30** en casual
    y en la **31** en japonés; el `'P&L Mensual'` de `plan-financiero` tiene
    `TOTAL COSTES FIJOS` en la **28** (casual) y en la **29** (japonés).
  · `calculadora-ticket-medio` tiene **tres** rejillas distintas: 3 escenarios
    en columnas B/C/D (representante), una sola columna B con pares %/precio
    (5 hermanos) y una tabla de **mix de producto** con `PVP €` / `Mix % ventas`
    / `Aporte €` (panadería).
  · `cash-flow-break-even` tiene **tres**: `Cash Flow 12 Meses` con columna de
    total (representante), `Cash Flow` + `Break-Even` (hermanos) y
    `Cash Flow 24m` (panadería).
  · `calculadora-capex` tiene **tres**: rangos bajo/medio/alto (representante),
    `CAPEX Desglosado` con `#`/`Categoría`/`Partida` (hermanos) e
    `Importe Mínimo`/`Importe Máximo` (panadería).

Por eso cada fichero se resuelve en dos pasos: **(1) detectar la variante por
firma estructural** —y **abortar** con `VarianteDesconocida` si no encaja, nunca
«aplica la del representante»— y **(2) localizar cada fila por su ETIQUETA**
(`_fila()`, sin acentos y por regex). Lo que cambia de guía a guía —cifras de
ejemplo, mapa de conceptos, notas— vive en `contenido_<pid>/a.py` y llega por el
parámetro `contenido`; este módulo no teclea ni un importe del oficio.

Qué implementa, por id de la SPEC:
  §2.1  TEC-01/DOM-08/COM-04 — el ticket medio se calcula (ponderado por mix),
        con fila de control «% de comensales asignado» y facturación día/mes.
  §2.2  TEC-02/DOM-07/COM-05 + §7-bis.14 — P&L de 3 escenarios ENCADENADO, en
        las tres variantes, con el «Pesimista» recalibrado (malo, no inviable).
  §2.3  TEC-07/08/09/10, DOM-06/22/26/27, COM-07 — hoja «Proyección 3 Años»
        nueva, EBITDA que NO resta amortización + EBIT, hoja «Financiación» con
        cuadro francés (anualidad algebraica: pycel no implementa `PMT`), fondo
        de maniobra dimensionado (≥ 6 meses) y la correspondencia 22→12 del
        CAPEX (§2.3.6).
  §2.4  TEC-03/DOM-09/COM-06 — totales, flujo neto, acumulado, las tres filas
        de IVA del modelo 303, la cuota del préstamo y el break-even en meses
        y en €.
  §2.5  NUEVO-01 (el `P&L Mensual` de los 5 hermanos sin ni un total) y
        NUEVO-02 (panadería: EBITDA = facturación, margen 122,97 %).
  §7-bis.13 — «sin dato» se escribe `""`, **nunca `0`**: con el libro en blanco
        ningún margen dice «0,0 %» ni se enciende un semáforo.

Lo que NO hace (§1.13/§7.1): no crea ficheros, no renombra, no escribe
`externalLink` entre libros (la cuota del préstamo se REPITE en el cash flow
como celda verde con la nota de dónde sale), no toca `paperSize` de las hojas
que ya existen —a las que CREA sí les pone el A4 completo, o `demo_a4` las
contaría como `noprint`— y no borra ninguna fila que el cliente pueda haber
rellenado.

⚠️ U+202F (espacio fino) y U+2011 (guion no separable) se referencian por
escape desde `motor.NARROW` / `motor.NOBRK`, nunca escribiendo el carácter.
"""
import re
import unicodedata

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

import motor
from motor import FMT_ENT, FMT_EUR, FMT_PCT

SPEC = 'guias-v2-SPEC.md §2'

FICHEROS = ['calculadora-ticket-medio.xlsx',
            'pl-mensual-escenarios.xlsx',
            'plan-financiero-3-anos.xlsx',
            'cash-flow-break-even.xlsx',
            'calculadora-capex.xlsx']

#: `motor.aplicar`/`motor.cerrar` se aplican a los cinco: el grupo sólo añade.
PROPIOS = []

#: `guia-dark-kitchen` no tiene ninguno de los cinco. Sin esta guarda,
#: `main.ficheros_a_tocar()` añadiría los cinco nombres a la lista de trabajo y
#: `procesar()` intentaría abrir un fichero que no existe.
SIN_GRUPO_A = ('guia-dark-kitchen',)


def ficheros(ctx):
    pid = (ctx or {}).get('producto')
    return [] if pid in SIN_GRUPO_A else list(FICHEROS)


#: §1.2/§7-bis.12 — cada constante que el grupo convierte en fórmula queda
#: anotada aquí `(fichero, hoja, celda, valor anterior, fórmula, nota)` y
#: `demos()` la RECALCULA con pycel: el valor nuevo tiene que coincidir con el
#: viejo con tolerancia 0,01 €, o la diferencia tiene que venir con una nota que
#: la justifique (el caso de NUEVO-02, donde el número viejo estaba mal sumado).
#: Sin esta comprobación, «conservar el número» es una promesa sin gate.
SUSTITUCIONES = {}


def a_formula(ws, coord, formula, fname, cambios, fmt=None, celda_ejemplo=None,
              nota_dif=None):
    anterior = motor.a_formula(ws, coord, formula, celda_ejemplo=celda_ejemplo,
                               fmt=fmt, informe=cambios, fname=fname,
                               nota=nota_dif)
    if anterior is not None and isinstance(anterior, (int, float)) \
            and not isinstance(anterior, bool):
        SUSTITUCIONES.setdefault(fname, []).append(
            {'hoja': ws.title, 'celda': coord, 'anterior': anterior,
             'formula': formula, 'nota': nota_dif})
    return anterior


class VarianteDesconocida(Exception):
    """§1.1/§7-bis.11 aplicado al grupo A: el módulo NO adivina una rejilla."""


# ==========================================================================
# Localización por ETIQUETA (nunca por posición fija)
# ==========================================================================
def _sin_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')


def _norm(v):
    """Etiqueta normalizada: sin acentos, sin dobles espacios, en minúscula.

    También normaliza el ESPACIO FINO (U+202F) y el GUION NO SEPARABLE (U+2011)
    de la familia a sus equivalentes normales: un `.md`/`.xlsx` de esta familia
    puede traerlos y un patrón escrito con espacio normal no encontraría nada.
    """
    if not isinstance(v, str):
        return ''
    # La convención tipográfica se aplica ANTES de comparar: el barrido
    # de `motor.normalizar_texto` cambia «≤» por «<=» y «sólo» por «solo»
    # en el fichero, y sin esto el módulo de contenido —que sigue
    # escribiéndolos a la vieja usanza— no reconocería su propia salida.
    t = motor.convencion(v).replace(motor.NARROW, ' ').replace(
        motor.NOBRK, '-')
    t = _sin_acentos(t).lower()
    return re.sub(r'\s+', ' ', t).strip()


def etiquetas(ws, col=1, desde=1, hasta=None):
    """[(fila, etiqueta normalizada, etiqueta cruda)] de una columna."""
    fuera = []
    tope = hasta or ws.max_row
    for r in range(desde, tope + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            fuera.append((r, _norm(v), v))
    return fuera


def _fila(ws, patron, col=1, desde=1, hasta=None, obligatoria=True,
          fname='', saltar=()):
    """Primera fila cuya etiqueta case con `patron` (regex, ya normalizado).

    Devuelve la PRIMERA coincidencia de arriba abajo: es lo que hace estable la
    2.ª pasada cuando el grupo ha añadido más abajo una etiqueta parecida (p. ej.
    `FLUJO DE CAJA NETO` en la fila 21 y `FLUJO DE CAJA NETO CON IVA Y DEUDA` en
    la 30). `saltar` excluye filas concretas por si la etiqueta nueva es un
    prefijo exacto de la vieja.
    """
    rx = re.compile(patron)
    for r, norm, _crudo in etiquetas(ws, col, desde, hasta):
        if r in saltar:
            continue
        if rx.search(norm):
            return r
    if obligatoria:
        raise VarianteDesconocida(
            (fname + ':' if fname else '') + ws.title + ': no hay ninguna fila '
            'cuya etiqueta case con ' + repr(patron) + ' en la columna '
            + str(col) + '. El grupo A NO escribe por posición fija: revisa la '
            'rejilla o añade el patrón en contenido_<pid>/a.py.')
    return None


RX_MES = re.compile(r'^(mes\s*\d+|m\d+|ene|feb|mar|abr|may|jun|jul|ago|sep|'
                    r'set|oct|nov|dic|enero|febrero|marzo|abril|mayo|junio|'
                    r'julio|agosto|septiembre|octubre|noviembre|diciembre)$')
RX_TOTAL_COL = re.compile(r'^(total|total anual|anual|ano \d|total ano)')


def columnas_mes(ws, fila_cab):
    """(columnas de mes, columna de total) leídas de la CABECERA.

    Representante: `Mes 1`…`Mes 12` + `Total Anual`. Hermanos: `Ene`…`Dic` sin
    columna de total. Panadería: `M1`…`M12` + `Año 1` + `Año 2`. Tres rejillas,
    una sola lectura.
    """
    from openpyxl.utils import get_column_letter as gcl
    meses, totales = [], []
    for c in range(2, ws.max_column + 1):
        cab = _norm(ws.cell(row=fila_cab, column=c).value)
        if RX_MES.match(cab):
            meses.append(gcl(c))
        elif RX_TOTAL_COL.match(cab):
            totales.append(gcl(c))
    return meses, (totales[0] if totales else None)


def _a4(ws):
    """A4 completo en una hoja NUEVA (§1.13: el motor protege las que ya
    existían, no las que no existen; sin esto `demo_a4` la cuenta `noprint`)."""
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8


def hoja(wb, nombre, tras=None):
    """Hoja `nombre`, creándola (con A4) si falta. Idempotente."""
    if nombre in wb.sheetnames:
        return wb[nombre], False
    indice = None
    if tras and tras in wb.sheetnames:
        indice = wb.sheetnames.index(tras) + 1
    ws = wb.create_sheet(nombre) if indice is None \
        else wb.create_sheet(nombre, indice)
    _a4(ws)
    return ws, True


def cabecera(ws, titulo, subtitulo, columnas, fila=4, ancho_a=44):
    """Título + subtítulo + fila de cabecera con el molde de la familia."""
    motor.val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=14)
    motor.val(ws, 'A2', subtitulo)
    ws['A2'].font = Font(italic=True, size=9)
    from openpyxl.utils import get_column_letter as gcl
    for i, texto in enumerate(columnas):
        cel = ws.cell(row=fila, column=i + 1, value=texto)
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor='2D2D2D')
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    ws.column_dimensions['A'].width = ancho_a
    for i in range(1, len(columnas)):
        L = gcl(i + 1)
        if not ws.column_dimensions[L].width:
            ws.column_dimensions[L].width = 16.0


def base_bloque(ws, marca_rx):
    """Fila donde empieza un bloque que el grupo AÑADE, de forma IDEMPOTENTE.

    Calcularla como `ws.max_row + 2` funciona la primera vez y **desplaza el
    bloque dos filas en cada pasada siguiente**: medido, 100 diferencias entre
    la 1.ª y la 2.ª pasada del cash flow. Igual que el `MARCA_BLOQUE` de
    `motor.cerrar_checklist`, la posición se ancla en la ETIQUETA del propio
    bloque: si ya está escrito se reescribe encima, y si no, va al final.
    """
    r = _fila(ws, marca_rx, obligatoria=False)
    return r if r else ws.max_row + 2


def nota(ws, coord, texto):
    """Línea de nota en cursiva pequeña (no es dato: no la mira ningún gate)."""
    motor.val(ws, coord, texto, wrap=True)
    ws[coord].font = Font(italic=True, size=9)


def apunta(cambios, fname, ws, texto):
    cambios.append(fname + ':' + ws.title + ': ' + texto)


# ==========================================================================
# §2.1 · calculadora-ticket-medio.xlsx
# ==========================================================================
#: Etiqueta de la fila de control que el grupo AÑADE (§2.1). Se guarda como
#: constante porque el detector de pares tiene que EXCLUIRLA en la 2.ª pasada:
#: empieza por «%» y, si no se saltara, se leería como un tramo de mix más.
ETIQUETA_MIX = '% de comensales asignado (debe sumar 100 %)'
RX_ETIQUETA_MIX = re.compile(r'^% de comensales asignado')

RX_PCT_FILA = re.compile(r'^%\s')
RX_PRECIO_FILA = re.compile(r'\(\s*(eur|€)\s*\)|precio|ticket')
RX_TICKET_RES = re.compile(r'^ticket medio (ponderado|estimado|proyectado)')


#: Los tres rótulos con los que puede venir cabecerada una columna de
#: escenario, ya normalizados (sin acentos y en minúscula).
NOMBRES_ESCENARIO = ('pesimista', 'realista', 'optimista', 'conservador',
                     'base', 'moderado')


def variante_ticket(wb, fname=''):
    """`'escenarios-3col'` | `'columna-unica'` | `'mix-producto'`.

    Firma estructural MEDIDA (2026-08-29), nunca el nombre de la guía:
      · 3 columnas de escenario en la cabecera (`Escenario 1/2/3`) → rep.
      · sin fila de cabecera y valores sólo en B → los 5 hermanos.
      · cabecera `PVP …` + `Mix % ventas` + `Aporte …` → panadería.
    """
    if 'Ticket Medio' not in wb.sheetnames:
        raise VarianteDesconocida(
            fname + ': no hay hoja «Ticket Medio». Hojas: '
            + repr(wb.sheetnames))
    ws = wb['Ticket Medio']
    for fila in (3, 4):
        cab = [_norm(ws.cell(row=fila, column=c).value)
               for c in range(1, ws.max_column + 1)]
        # El mix de producto se mira PRIMERO: su cabecera también tiene tres
        # columnas y si el otro criterio se ensancha lo atraparía antes.
        if any('mix' in c for c in cab) and any('aporte' in c for c in cab):
            return 'mix-producto', fila
        # RC-02 renombra «Escenario 1/2/3» a «Pesimista/Realista/Optimista»
        # para poder cruzarlos con el P&L. El detector NO puede depender de la
        # palabra «escenario»: en la 2.ª pasada dejaba de reconocer su propia
        # salida y el libro se procesaba como si fuera el de los hermanos —
        # 10 diferencias de idempotencia y la hoja reescrita con otra rejilla.
        if any(c.startswith('escenario ') or c in NOMBRES_ESCENARIO
               for c in cab):
            return 'escenarios-3col', fila
    if _fila(ws, RX_TICKET_RES.pattern, obligatoria=False):
        return 'columna-unica', None
    raise VarianteDesconocida(
        fname + ':' + ws.title + ': rejilla de ticket medio no reconocida. '
        'fila3=' + repr([ws.cell(row=3, column=c).value
                         for c in range(1, ws.max_column + 1)])
        + ' fila4=' + repr([ws.cell(row=4, column=c).value
                            for c in range(1, ws.max_column + 1)]))


def pares_mix(ws, desde, hasta):
    """[(fila_%, fila_precio)] + [filas de precio INCONDICIONAL] del bloque.

    Regla estructural, la misma para las tres rejillas: una fila cuya etiqueta
    empieza por «%» es un TRAMO del mix y su precio es la fila inmediatamente
    siguiente; una fila de precio que no viene precedida de un «%» la pide el
    100 % de los comensales y entra SIN ponderar. Es exactamente el caso que la
    SPEC describe en japonés (`B7` «Precio medio ramen/principal» sin `%`) y el
    del representante (cinco pares y ningún incondicional).
    """
    filas = dict((r, n) for r, n, _ in etiquetas(ws, 1, desde, hasta))
    pares, sueltos, consumidas = [], [], set()
    for r in sorted(filas):
        if r in consumidas or RX_ETIQUETA_MIX.match(filas[r]):
            continue
        if RX_PCT_FILA.match(filas[r]):
            siguiente = r + 1
            if siguiente in filas and RX_PRECIO_FILA.search(filas[siguiente]):
                pares.append((r, siguiente))
                consumidas.add(siguiente)
            continue
        if RX_PRECIO_FILA.search(filas[r]):
            sueltos.append(r)
    return pares, sueltos


def _formula_ticket(col, pares, sueltos):
    trozos = [col + str(p) + '*' + col + str(q) for p, q in pares]
    trozos += [col + str(r) for r in sueltos]
    return '=' + '+'.join(trozos)


def ticket_escenarios_3col(wb, fname, cambios, contenido):
    """Representante: tres escenarios en B/C/D (TEC-01, DOM-08, COM-04)."""
    ws = wb['Ticket Medio']
    cols = [c for c in ('B', 'C', 'D')
            if (_norm(ws[c + '4'].value).startswith('escenario ')
                or _norm(ws[c + '4'].value) in NOMBRES_ESCENARIO)]
    f_ticket = _fila(ws, RX_TICKET_RES.pattern, fname=fname)
    f_cub = _fila(ws, r'^cubiertos\s*/\s*d[ií]a|^cubiertos por dia',
                  hasta=ws.max_row, fname=fname)
    f_fdia = _fila(ws, r'^facturacion diaria', fname=fname)
    f_dias = _fila(ws, r'^dias abierto', fname=fname)
    f_fmes = _fila(ws, r'^facturacion mensual', fname=fname)
    pares, sueltos = pares_mix(ws, 5, f_ticket - 1)
    if not pares:
        raise VarianteDesconocida(
            fname + ':' + ws.title + ': 0 pares %/precio entre las filas 5 y '
            + str(f_ticket - 1) + ': la rejilla no es la que este código sabe '
            'leer.')

    # fila de control del mix, en el hueco que ya existe sobre el resultado
    f_mix = f_ticket - 1
    if _norm(ws['A' + str(f_mix)].value):
        f_mix = None                      # no hay hueco: no se pisa contenido
    conf = (contenido.TICKET if contenido and hasattr(contenido, 'TICKET')
            else {})
    tramos = conf.get('filas_mix')        # regex de las filas que suman 100 %
    if f_mix and tramos:
        filas_mix = [p for p, _q in pares
                     if any(re.search(t, _norm(ws['A' + str(p)].value))
                            for t in tramos)]
        if filas_mix:
            motor.val(ws, 'A' + str(f_mix), ETIQUETA_MIX)
            for col in cols:
                motor.f(ws, col + str(f_mix),
                        '=' + '+'.join(col + str(p) for p in filas_mix),
                        fmt=FMT_PCT)
            motor.regla_expresion(
                ws, cols[0] + str(f_mix) + ':' + cols[-1] + str(f_mix),
                '=AND(ISNUMBER(' + cols[0] + str(f_mix) + '),'
                + cols[0] + str(f_mix) + '<>1)')
            nota(ws, 'E' + str(f_mix),
                 'Los tramos del menú se reparten el 100 % de los comensales. '
                 'El maridaje y la copa NO entran aquí: son consumo adicional '
                 'sobre el mismo comensal, por eso pueden sumar más de 100 %.')
            apunta(cambios, fname, ws, 'fila de control del mix en A'
                   + str(f_mix) + ' (§2.1) sobre ' + str(len(filas_mix))
                   + ' tramos')

    for col in cols:
        motor.f(ws, col + str(f_ticket), _formula_ticket(col, pares, sueltos),
                fmt=FMT_EUR, bold=True)
        motor.f(ws, col + str(f_fdia),
                '=' + col + str(f_ticket) + '*' + col + str(f_cub),
                fmt=FMT_EUR)
        motor.f(ws, col + str(f_fmes),
                '=' + col + str(f_fdia) + '*' + col + str(f_dias),
                fmt=FMT_EUR)
    # el verde se retira SÓLO de las filas que ahora calcula el libro: las de
    # `Cubiertos/día` y `Días abierto/mes` siguen siendo entradas del cliente
    # (TEC-01 pinta de verde las cinco, incluidas las tres de resultado).
    for r in (f_ticket, f_fdia, f_fmes):
        motor.quitar_verde(ws, cols[0] + str(r) + ':' + cols[-1] + str(r))
    # RC-02 · las columnas se llamaban «Escenario 1/2/3» y el P&L de escenarios
    # «Pesimista/Realista/Optimista»: el comprador no podía cruzarlos, y el
    # «escenario 3» tenía MENOS cubiertos que el 2.
    for col, titulo in (conf.get('cabeceras') or {}).items():
        if col in cols:
            motor.val(ws, col + '4', titulo, bold=True)
    apunta(cambios, fname, ws, str(len(pares)) + ' pares %/precio + '
           + str(len(sueltos)) + ' incondicionales → ticket ponderado en la '
           'fila ' + str(f_ticket) + ', facturación día/mes (TEC-01, DOM-08)')
    _precargar_ticket(ws, cols, conf, cambios, fname)
    _ticket_iva_y_cuadre(ws, cols, conf, cambios, fname, f_ticket, f_fmes)


def _ticket_iva_y_cuadre(ws, cols, conf, cambios, fname, f_ticket, f_fmes):
    """RD-23 + RC-02 — la hoja FIJA precios y no decía si llevan IVA, y su
    facturación no se podía cruzar con la del P&L de escenarios.

    §1.5(a) obliga a que toda hoja que fije un PVP lleve el precio sin IVA y
    con IVA, con el tipo en celda: en España la carta se anuncia CON IVA y un
    menú de 150 € son 136,36 € netos. Y §7-bis.7 exige una sola cifra por
    magnitud: aquí se repite la del P&L con su nota (§1.13, no `externalLink`)
    y se marca en rojo cualquier desvío mayor del 2 %.
    """
    base = base_bloque(ws, r'^iva y cuadre con el p&l')
    iva = conf.get('iva') or {}
    r = base
    motor.val(ws, 'A' + str(r), 'IVA Y CUADRE CON EL P&L DE ESCENARIOS',
              bold=True)
    r += 1
    motor.val(ws, 'A' + str(r),
              iva.get('etiqueta', 'Tipo de IVA de los precios (%)'))
    motor.val(ws, 'B' + str(r), iva.get('valor', 0.10), fmt=FMT_PCT,
              verde_=True)
    motor.fijar_formato(ws, 'B' + str(r), FMT_PCT)
    celda_iva = '$B$' + str(r)
    nota(ws, 'C' + str(r), iva.get('nota', ''))
    r += 1
    motor.val(ws, 'A' + str(r), 'Ticket medio CON IVA — el precio que ve el '
              'comensal (€)')
    for col in cols:
        motor.f(ws, col + str(r),
                '=IF(' + col + str(f_ticket) + '="","",' + col + str(f_ticket)
                + '*(1+' + celda_iva + '))', fmt=FMT_EUR)
    r += 1
    ref = conf.get('reconciliacion') or {}
    if ref:
        motor.val(ws, 'A' + str(r), 'Facturación mensual del P&L de escenarios '
                  '(€) — la que va al banco')
        for col in cols:
            motor.val(ws, col + str(r), ref.get(col), fmt=FMT_EUR)
        f_ref = r
        r += 1
        motor.val(ws, 'A' + str(r), 'Diferencia con este simulador (%)')
        for col in cols:
            motor.f(ws, col + str(r),
                    '=IF(OR(' + col + str(f_ref) + '="",' + col + str(f_ref)
                    + '=0,' + col + str(f_fmes) + '=""),"",(' + col
                    + str(f_fmes) + '-' + col + str(f_ref) + ')/' + col
                    + str(f_ref) + ')', fmt=FMT_PCT)
        motor.semaforo_isnumber(ws, cols[0] + str(r) + ':' + cols[-1] + str(r),
                                cols[0] + str(r), operador='>', umbral='0.02')
        r += 1
        nota(ws, 'A' + str(r),
             'Las dos hojas describen el MISMO restaurante: mismos cubiertos/'
             'día y mismos días abiertos. Los tickets de comida y cena del P&L '
             'están calibrados contra el ticket ponderado de esta hoja, así '
             'que esta fila tiene que dar 0,0 %. Si no lo da, has cambiado un '
             'precio aquí y no allí — y la cifra que vale es la del P&L '
             '(§7-bis.7).')
    apunta(cambios, fname, ws, 'bloque de IVA (§1.5a) y cuadre con el P&L en '
           'las filas ' + str(base) + '-' + str(r) + ' (RD-23, RC-02)')
    return base


def _precargar_ticket(ws, cols, conf, cambios, fname):
    """§7-bis.7 — cifras de ejemplo que el propio producto ya defiende.

    Van en celda VERDE (editables) y salen de `contenido_<pid>/a.py`, con la
    fuente anotada allí. Si el módulo de contenido no las trae, la hoja queda
    con sus fórmulas y sin ejemplo: no se inventa ninguna cifra aquí.
    """
    escenarios = (conf or {}).get('escenarios')
    if not escenarios:
        return
    puestas = 0
    for col in cols:
        datos = escenarios.get(col)
        if not datos:
            continue
        for patron, valor in datos.items():
            r = _fila(ws, patron, obligatoria=False)
            if r is None:
                continue
            cel = ws[col + str(r)]
            if cel.data_type == 'f':
                continue
            motor.val(ws, col + str(r), valor, verde_=True)
            puestas += 1
    if puestas:
        apunta(cambios, fname, ws, str(puestas) + ' celdas de ejemplo '
               'precargadas en verde (§7-bis.7)')


def ticket_columna_unica(wb, fname, cambios, contenido):
    """Los 5 hermanos: una sola columna B, con `TICKET MEDIO ESTIMADO` vacío.

    Además de la fórmula que falta, se AÑADEN debajo `Cubiertos/día`,
    `Días abierto/mes`, `Facturación diaria` y `Facturación mensual`, que hoy no
    existen en el hermano y sí en el representante (§2.1).
    """
    ws = wb['Ticket Medio']
    f_ticket = _fila(ws, RX_TICKET_RES.pattern, fname=fname)
    pares, sueltos = pares_mix(ws, 4, f_ticket - 1)
    if not pares:
        raise VarianteDesconocida(
            fname + ':' + ws.title + ': 0 pares %/precio antes de la fila '
            + str(f_ticket))
    motor.f(ws, 'B' + str(f_ticket), _formula_ticket('B', pares, sueltos),
            fmt=FMT_EUR, bold=True)
    motor.quitar_verde(ws, 'B' + str(f_ticket))
    conf = (contenido.TICKET if contenido and hasattr(contenido, 'TICKET')
            else {})
    base = f_ticket + 2
    filas = (('Cubiertos/día', conf.get('cubiertos_dia'), FMT_ENT, True),
             ('Días abierto/mes', conf.get('dias_mes'), FMT_ENT, True),
             ('Facturación diaria (€)', None, FMT_EUR, False),
             ('Facturación mensual (€)', None, FMT_EUR, False))
    for i, (etiqueta, valor, fmt, editable) in enumerate(filas):
        r = base + i
        motor.val(ws, 'A' + str(r), etiqueta)
        if editable:
            motor.val(ws, 'B' + str(r), valor, fmt=fmt,
                      verde_=valor is not None)
            if valor is None:
                motor.verde(ws, 'B' + str(r))
    motor.f(ws, 'B' + str(base + 2),
            '=B' + str(f_ticket) + '*B' + str(base), fmt=FMT_EUR)
    motor.f(ws, 'B' + str(base + 3),
            '=B' + str(base + 2) + '*B' + str(base + 1), fmt=FMT_EUR)
    apunta(cambios, fname, ws, 'ticket ponderado en B' + str(f_ticket)
           + ' (' + str(len(pares)) + ' pares, ' + str(len(sueltos))
           + ' incondicionales) + cubiertos/día, días/mes y facturación '
             'día/mes en A' + str(base) + ':B' + str(base + 3) + ' (§2.1)')


def ticket_mix_producto(wb, fname, cambios, contenido, fila_cab):
    """Panadería: `PVP €` × `Mix % ventas` / 100 → `Aporte €`, ya calculado.

    Lo que falta es el CONTROL: la columna de mix suma 100 y nadie lo comprueba,
    y el ticket resultante (3,77 €) no cuadra con el 5,80 € que usa el P&L de la
    misma guía. Se añade el semáforo del mix y las filas de facturación.
    """
    ws = wb['Ticket Medio']
    f_tot = _fila(ws, r'^ticket medio', fname=fname)
    from openpyxl.utils import get_column_letter as gcl
    col_mix = col_aporte = None
    for c in range(2, ws.max_column + 1):
        cab = _norm(ws.cell(row=fila_cab, column=c).value)
        if 'mix' in cab:
            col_mix = gcl(c)
        elif 'aporte' in cab:
            col_aporte = gcl(c)
    if not (col_mix and col_aporte):
        raise VarianteDesconocida(fname + ':' + ws.title
                                  + ': falta la columna de mix o la de aporte')
    ultima = f_tot - 1
    while ultima > fila_cab and ws['A' + str(ultima)].value is None:
        ultima -= 1
    motor.f(ws, col_mix + str(f_tot),
            '=SUM(' + col_mix + str(fila_cab + 1) + ':' + col_mix
            + str(ultima) + ')', fmt=FMT_ENT, bold=True)
    motor.f(ws, col_aporte + str(f_tot),
            '=SUM(' + col_aporte + str(fila_cab + 1) + ':' + col_aporte
            + str(ultima) + ')', fmt=FMT_EUR, bold=True)
    motor.regla_expresion(
        ws, col_mix + str(f_tot),
        '=AND(ISNUMBER(' + col_mix + str(f_tot) + '),' + col_mix + str(f_tot)
        + '<>100)')
    apunta(cambios, fname, ws, 'mix y aporte totalizados en la fila '
           + str(f_tot) + ' con semáforo si el mix no suma 100 (§2.1)')


def ticket_medio(wb, fname, cambios, contenido):
    variante, fila_cab = variante_ticket(wb, fname)
    if variante == 'escenarios-3col':
        ticket_escenarios_3col(wb, fname, cambios, contenido)
    elif variante == 'columna-unica':
        ticket_columna_unica(wb, fname, cambios, contenido)
    else:
        ticket_mix_producto(wb, fname, cambios, contenido, fila_cab)
    return variante


# ==========================================================================
# §2.2 · pl-mensual-escenarios.xlsx  (TEC-02, DOM-07, COM-05, §7-bis.14)
# ==========================================================================
#: Patrones de fila, normalizados. Sirven para las TRES variantes porque los
#: tres generadores heredaron los mismos rótulos: lo que cambia es la fila.
RX_TOT_INGRESOS = r'^total ingresos|^facturacion total mensual'
RX_FOOD_COST = r'^food cost|^materia prima|^coste materias primas$'
RX_TOT_VARIABLES = r'^total costes variables'
RX_TOT_FIJOS = r'^total costes fijos'
RX_EBITDA = r'^ebitda(?! anual)'
RX_PCT_EBITDA = r'^% ebitda|^margen ebitda'
RX_BLOQUE_ING = r'^ingresos$'
RX_BLOQUE_VAR = r'^costes variables$'
RX_BLOQUE_FIJ = r'^costes fijos$'


def _ultimo_detalle(ws, desde, hasta, saltar_rx):
    """Última fila con etiqueta entre `desde` y `hasta` que no sea un total."""
    ultima = None
    for r, norm, _c in etiquetas(ws, 1, desde, hasta):
        if any(re.search(rx, norm) for rx in saltar_rx):
            continue
        ultima = r
    return ultima


def pl_escenarios_columnas(wb, fname, cambios, contenido):
    """Representante: una hoja `Escenarios`, tres columnas B/C/D.

    Hoy: 0 fórmulas, 0 valores y `B17:D22` en VERDE, es decir, la hoja le pide
    al cliente que teclee él el EBITDA (TEC-02).
    """
    ws = wb['Escenarios']
    cols = [c for c in ('B', 'C', 'D')
            if isinstance(ws[c + '4'].value, str)]
    f_cub_com = _fila(ws, r'^cubiertos.*comida', fname=fname)
    f_cub_cena = _fila(ws, r'^cubiertos.*cena', fname=fname)
    f_tk_com = _fila(ws, r'^ticket medio comida', fname=fname)
    f_tk_cena = _fila(ws, r'^ticket medio cena', fname=fname)
    f_dias = _fila(ws, r'^dias abierto', fname=fname)
    f_fc = _fila(ws, r'^food cost', fname=fname)
    f_pers = _fila(ws, r'^coste personal', fname=fname)
    f_res = _fila(ws, r'^resultados$', fname=fname)
    f_otros = _ultimo_detalle(ws, f_pers, f_res - 1, ())
    f_fact = _fila(ws, r'^facturacion mensual', desde=f_res, fname=fname)
    f_mp = _fila(ws, r'^coste materia prima', desde=f_res, fname=fname)
    f_mb = _fila(ws, r'^margen bruto', desde=f_res, fname=fname)
    f_cf = _fila(ws, r'^costes fijos totales', desde=f_res, fname=fname)
    f_eb = _fila(ws, RX_EBITDA, desde=f_res, fname=fname)
    f_me = _fila(ws, RX_PCT_EBITDA, desde=f_res, fname=fname)

    for col in cols:
        motor.f(ws, col + str(f_fact),
                '=(' + col + str(f_cub_com) + '*' + col + str(f_tk_com) + '+'
                + col + str(f_cub_cena) + '*' + col + str(f_tk_cena) + ')*'
                + col + str(f_dias), fmt=FMT_EUR)
        motor.f(ws, col + str(f_mp),
                '=' + col + str(f_fact) + '*' + col + str(f_fc), fmt=FMT_EUR)
        motor.f(ws, col + str(f_mb),
                '=' + col + str(f_fact) + '-' + col + str(f_mp), fmt=FMT_EUR)
        motor.f(ws, col + str(f_cf),
                '=SUM(' + col + str(f_pers) + ':' + col + str(f_otros) + ')',
                fmt=FMT_EUR)
        motor.f(ws, col + str(f_eb),
                '=' + col + str(f_mb) + '-' + col + str(f_cf), fmt=FMT_EUR,
                bold=True)
        # §7-bis.13: un mes sin una sola venta NO tiene un margen del «0,0 %»
        motor.f(ws, col + str(f_me),
                '=IF(' + col + str(f_fact) + '=0,"",' + col + str(f_eb) + '/'
                + col + str(f_fact) + ')', fmt=FMT_PCT)
        motor.quitar_verde(ws, col + str(f_fact) + ':' + col + str(f_me))
    # RD-11 · había DOS EBITDA distintos en el mismo producto: plan-financiero
    # excluye la amortización (TEC-08) y esta hoja la metía dentro de «Otros
    # costes fijos» por decisión declarada en su propia nota, con la fila
    # rotulada «EBITDA mensual». Eso es un EBIT con otro nombre.
    f_amo, f_ebit = f_me + 1, f_me + 2
    motor.val(ws, 'A' + str(f_amo), 'Amortización mensual (€)')
    motor.val(ws, 'A' + str(f_ebit), 'EBIT — resultado de explotación (€)')
    for col in cols:
        motor.verde(ws, col + str(f_amo))
        motor.fijar_formato(ws, col + str(f_amo), FMT_EUR)
        motor.f(ws, col + str(f_ebit),
                '=IF(' + col + str(f_eb) + '="","",' + col + str(f_eb) + '-IF('
                'ISNUMBER(' + col + str(f_amo) + '),' + col + str(f_amo)
                + ',0))', fmt=FMT_EUR, bold=True)
    motor.permitir_negativo(ws, cols[0] + str(f_ebit) + ':' + cols[-1]
                            + str(f_ebit))
    _semaforo_negativo(ws, cols, (f_eb, f_me, f_ebit))
    apunta(cambios, fname, ws, 'P&L encadenado en las filas ' + str(f_fact)
           + '-' + str(f_me) + ' para ' + str(len(cols))
           + ' escenarios; verde retirado de los resultados (TEC-02, DOM-07); '
           'amortización FUERA del EBITDA en la fila ' + str(f_amo)
           + ' y EBIT en la ' + str(f_ebit) + ' (RD-11)')
    _precargar_pl(ws, cols, contenido, cambios, fname)


def _semaforo_negativo(ws, cols, filas):
    """§1.6 — rojo si el resultado es negativo, con la guarda `ISNUMBER`.

    Sin ella Excel evalúa `""<0` como FALSO y cualquier TEXTO como VERDADERO:
    el semáforo pintaría de rojo justo la celda que dice que no hay dato.
    """
    for r in filas:
        rango = cols[0] + str(r) + ':' + cols[-1] + str(r)
        motor.semaforo_isnumber(ws, rango, '$' + cols[0] + '$' + str(r)
                                if len(cols) == 1 else cols[0] + str(r))


def _precargar_pl(ws, cols, contenido, cambios, fname, clave='PL'):
    conf = getattr(contenido, clave, None) if contenido else None
    if not conf or not conf.get('escenarios'):
        return
    puestas = 0
    for col in cols:
        for patron, valor in (conf['escenarios'].get(col) or {}).items():
            r = _fila(ws, patron, obligatoria=False)
            if r is None or ws[col + str(r)].data_type == 'f':
                continue
            motor.val(ws, col + str(r), valor, verde_=True)
            puestas += 1
    for patron, texto in (conf.get('notas') or {}).items():
        r = _fila(ws, patron, obligatoria=False)
        if r is not None:
            nota(ws, 'E' + str(r), texto)
    if puestas:
        apunta(cambios, fname, ws, str(puestas) + ' celdas de ejemplo en verde '
               '(§7-bis.7/§7-bis.14: el Pesimista es un escenario malo, no '
               'inviable)')


def pl_tres_hojas(wb, fname, cambios, contenido):
    """Los 5 hermanos: `Pesimista`/`Realista`/`Optimista` con TODO tecleado.

    §1.2/§7-bis.12: las líneas de detalle se quedan VERDES con su valor actual
    como ejemplo y los TOTALES pasan a fórmula. El food cost sale del rótulo
    («Food cost (33%)») y baja a una celda verde: hoy va multiplicado a mano.
    La columna `D % s/Ventas` está rotulada y **no existe**.
    """
    conf = getattr(contenido, 'PL', None) if contenido else None
    for nombre in ('Pesimista', 'Realista', 'Optimista'):
        ws = wb[nombre]
        f_ing = _fila(ws, RX_BLOQUE_ING, fname=fname)
        f_tot_ing = _fila(ws, RX_TOT_INGRESOS, fname=fname)
        f_var = _fila(ws, RX_BLOQUE_VAR, fname=fname)
        f_fc = _fila(ws, RX_FOOD_COST, desde=f_var, fname=fname)
        f_tot_var = _fila(ws, RX_TOT_VARIABLES, fname=fname)
        f_fij = _fila(ws, RX_BLOQUE_FIJ, fname=fname)
        f_tot_fij = _fila(ws, RX_TOT_FIJOS, fname=fname)
        f_eb = _fila(ws, RX_EBITDA, desde=f_tot_fij, fname=fname)
        f_pct = _fila(ws, RX_PCT_EBITDA, desde=f_eb, obligatoria=False)
        cols = ['B']
        col_anual = 'C' if _norm(ws['C4'].value).startswith('anual') else None
        col_pct = 'D' if 's/ventas' in _norm(ws['D4'].value) else None

        motor.verde(ws, 'B' + str(f_ing + 1) + ':B' + str(f_tot_ing - 1))
        motor.verde(ws, 'B' + str(f_fij + 1) + ':B' + str(f_tot_fij - 1))
        a_formula(ws, 'B' + str(f_tot_ing),
                  '=SUM(B' + str(f_ing + 1) + ':B' + str(f_tot_ing - 1) + ')',
                  fname, cambios, fmt=FMT_EUR)
        # el food cost deja de ir en el rótulo y baja a celda verde
        celda_fc = _celda_food_cost(ws, f_fc, conf, cambios, fname)
        a_formula(ws, 'B' + str(f_fc), '=B' + str(f_tot_ing) + '*' + celda_fc,
                  fname, cambios, fmt=FMT_EUR)
        motor.verde(ws, 'B' + str(f_fc + 1) + ':B' + str(f_tot_var - 1))
        a_formula(ws, 'B' + str(f_tot_var),
                  '=SUM(B' + str(f_fc) + ':B' + str(f_tot_var - 1) + ')',
                  fname, cambios, fmt=FMT_EUR)
        a_formula(ws, 'B' + str(f_tot_fij),
                  '=SUM(B' + str(f_fij + 1) + ':B' + str(f_tot_fij - 1) + ')',
                  fname, cambios, fmt=FMT_EUR)
        a_formula(ws, 'B' + str(f_eb),
                  '=B' + str(f_tot_ing) + '-B' + str(f_tot_var) + '-B'
                  + str(f_tot_fij), fname, cambios, fmt=FMT_EUR)
        if f_pct:
            motor.f(ws, 'B' + str(f_pct),
                    '=IF(B' + str(f_tot_ing) + '=0,"",B' + str(f_eb) + '/B'
                    + str(f_tot_ing) + ')', fmt=FMT_PCT)
        if col_anual:
            for r, _n, _c in etiquetas(ws, 1, f_ing, f_eb):
                if ws['B' + str(r)].value is None:
                    continue
                a_formula(ws, col_anual + str(r), '=B' + str(r) + '*12',
                          fname, cambios, fmt=FMT_EUR)
            cols.append(col_anual)
        if col_pct:
            for r, _n, _c in etiquetas(ws, 1, f_ing, f_eb):
                if ws['B' + str(r)].value is None:
                    continue
                motor.f(ws, col_pct + str(r),
                        '=IF($B$' + str(f_tot_ing) + '=0,"",B' + str(r)
                        + '/$B$' + str(f_tot_ing) + ')', fmt=FMT_PCT)
        _semaforo_negativo(ws, ['B'], [f_eb] + ([f_pct] if f_pct else []))
        apunta(cambios, fname, ws, 'totales y % s/ventas encadenados; food '
               'cost en ' + celda_fc + ' (§2.2 variante tres hojas)')
        if conf and conf.get('escenarios', {}).get(nombre):
            _precargar_hoja(ws, nombre, conf, cambios, fname)


def _celda_food_cost(ws, f_fc, conf, cambios, fname):
    """Saca el food cost del rótulo «Food cost (33%)» y lo pone en celda verde.

    Se escribe en la columna E de la propia fila (libre en las tres hojas
    medidas). El porcentaje se LEE del rótulo: es el dato que el generador dejó
    ahí, no una cifra inventada.
    """
    crudo = ws['A' + str(f_fc)].value or ''
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', crudo)
    pct = float(m.group(1).replace(',', '.')) / 100 if m else None
    if (conf or {}).get('food_cost') is not None:
        pct = conf['food_cost']
    destino = 'E' + str(f_fc)
    if ws[destino].value is None and pct is not None:
        motor.val(ws, destino, pct, fmt=FMT_PCT, verde_=True)
        nota(ws, 'F' + str(f_fc),
             'Food cost objetivo. Estaba escrito dentro del rótulo y '
             'multiplicado a mano: ahora se cambia aquí y se recalcula.')
        apunta(cambios, fname, ws, 'food cost ' + str(pct) + ' extraído del '
               'rótulo ' + repr(crudo[:40]) + ' a la celda verde ' + destino)
    return '$' + destino[0] + '$' + destino[1:]


def _precargar_hoja(ws, nombre, conf, cambios, fname):
    puestas = 0
    for patron, valor in (conf['escenarios'][nombre] or {}).items():
        r = _fila(ws, patron, obligatoria=False)
        if r is None or ws['B' + str(r)].data_type == 'f':
            continue
        motor.val(ws, 'B' + str(r), valor, verde_=True)
        puestas += 1
    if puestas:
        apunta(cambios, fname, ws, str(puestas)
               + ' líneas recalibradas (§7-bis.14)')


def pl_hoja_unica(wb, fname, cambios, contenido):
    """Panadería: `P&L 3 escenarios`, que YA calcula. Sólo §1 más el margen.

    Se le añade la fila de margen EBITDA (%), que no tiene, y el semáforo.
    """
    ws = wb['P&L 3 escenarios']
    f_fact = _fila(ws, r'^facturacion total mensual', fname=fname)
    f_eb = _fila(ws, r'^ebitda mensual', fname=fname)
    cols = [c for c in ('B', 'C', 'D')]
    f_pct = _fila(ws, r'^margen ebitda', obligatoria=False)
    if f_pct is None:
        f_pct = ws.max_row + 1
        motor.val(ws, 'A' + str(f_pct), 'Margen EBITDA (%)')
    for col in cols:
        motor.f(ws, col + str(f_pct),
                '=IF(' + col + str(f_fact) + '=0,"",' + col + str(f_eb) + '/'
                + col + str(f_fact) + ')', fmt=FMT_PCT)
    _semaforo_negativo(ws, cols, (f_eb, f_pct))
    apunta(cambios, fname, ws, 'margen EBITDA (%) en la fila ' + str(f_pct)
           + ' + semáforo ISNUMBER (§2.2 variante hoja única)')


def pl_mensual(wb, fname, cambios, contenido):
    variante = motor.variante_pl(wb)
    if variante == 'escenarios-columnas':
        pl_escenarios_columnas(wb, fname, cambios, contenido)
    elif variante == 'tres-hojas':
        pl_tres_hojas(wb, fname, cambios, contenido)
    elif variante == 'hoja-unica':
        pl_hoja_unica(wb, fname, cambios, contenido)
    else:
        raise VarianteDesconocida(
            fname + ': `motor.variante_pl` no reconoce las hojas '
            + repr(wb.sheetnames))
    return variante


# ==========================================================================
# §2.3 · plan-financiero-3-anos.xlsx
# ==========================================================================
HOJA_PROY = 'Proyección 3 Años'
HOJA_FIN = 'Financiación'
AVISO_CARENCIA = 'La carencia no puede igualar ni superar el plazo'
ANOS_CUADRO = 10          # filas del cuadro francés; la nota explica ampliarlo

#: B-01 · las dos celdas verdes que reconcilian la proyección con el cash flow.
#: Las dos etiquetas DECLARAN su tipo (`meses` → recuento, `%` inicial →
#: porcentaje) porque `motor.tipo_por_etiqueta` decide el formato por el texto:
#: una etiqueta muda acaba en `#,##0.00 €` en la 2.ª pasada, que es exactamente
#: lo que rompió B-04 en panadería.
ET_RAMPA_MESES = 'Meses de rampa hasta el mes de crucero'
ET_RAMPA_PRIMERO = '% del crucero que se factura el primer mes'


def _rampa_ano_1(contenido, conf):
    """Los dos parámetros de la rampa del año 1, MEDIDOS, no inventados.

    B-01: `plan-financiero-3-anos` publicaba el año 1 como «mes tipo × 12»
    (2.276.736,00 €) y `cash-flow-break-even` como una rampa de arranque
    (2.049.062,40 €), 227.673,60 € aparte y sin una sola nota que reconciliara
    las dos cifras — el número de portada de los dos ficheros que el cliente
    lleva al banco.

    La rampa del cash flow ya está medida en `contenido_<pid>/a.py:CASH`, así
    que aquí no se elige ningún número: se DERIVA de ella.

    · `meses` = cuántos de los doce factores están por debajo del crucero.
    · `primero` = el porcentaje del primer mes que hace que la suma de una
      rampa LINEAL de `meses` tramos valga exactamente lo mismo que la del
      cash flow. Para la rampa del representante (0,60 · 0,70 · 0,80 · 0,85 ·
      0,90 · 0,95 y crucero) sale 0,60 clavado, porque su media es (p+1)/2.

    Si la guía no trae rampa medida, se devuelven 0 meses: la proyección se
    comporta EXACTAMENTE como antes (`P&L × 12`) y no se inventa un arranque.
    """
    conf_r = conf.get('rampa') if isinstance(conf, dict) else None
    if isinstance(conf_r, (list, tuple)) and len(conf_r) == 2:
        return int(conf_r[0]), float(conf_r[1]), None
    rampa = ((getattr(contenido, 'CASH', None) or {}).get('rampa')
             if contenido else None)
    if not rampa or len(rampa) != 12:
        return 0, 0.6, None
    meses = sum(1 for x in rampa if x < 1)
    suma = round(sum(rampa), 10)
    if meses <= 1:
        return 0, 0.6, suma
    primero = round(2 * (suma - (12 - meses)) / meses - 1, 6)
    if not 0 <= primero <= 1:
        # Una rampa que no se puede describir con dos parámetros no se fuerza:
        # se deja la proyección en «× 12» y la nota lo dice.
        return 0, 0.6, suma
    return meses, primero, suma


def variante_plan(wb, fname=''):
    """`'inversion-conceptos'` | `'inversion-categorias'` | `'pl-3-anos'`."""
    hojas = set(wb.sheetnames)
    if 'P&L 3 años' in hojas:
        return 'pl-3-anos'
    if 'Inversión' in hojas and 'P&L Mensual' in hojas:
        cab = _norm(wb['Inversión']['B4'].value)
        if cab == 'concepto':
            return 'inversion-conceptos'
        if cab == 'partida':
            return 'inversion-categorias'
    raise VarianteDesconocida(
        fname + ': plan financiero no reconocido. Hojas=' + repr(wb.sheetnames)
        + ' Inversión!B4=' + repr(wb['Inversión']['B4'].value
                                  if 'Inversión' in hojas else None))


def _celda(col, fila, fijo=False):
    return ('$' + col + '$' + str(fila)) if fijo else (col + str(fila))


# --------------------------------------------------------------------------
# 2.3.2 · el EBITDA deja de restar la amortización (TEC-08, DOM-26)
# --------------------------------------------------------------------------
def _pl_mensual_representante(wb, fname, cambios, contenido):
    ws = wb['P&L Mensual']
    f_tot_ing = _fila(ws, RX_TOT_INGRESOS, fname=fname)
    f_ing_bloque = _fila(ws, RX_BLOQUE_ING, fname=fname)
    f_tot_var = _fila(ws, RX_TOT_VARIABLES, fname=fname)
    f_var_bloque = _fila(ws, RX_BLOQUE_VAR, fname=fname)
    f_mb = _fila(ws, r'^margen bruto', fname=fname)
    f_fij = _fila(ws, RX_BLOQUE_FIJ, fname=fname)
    f_tot_fij = _fila(ws, RX_TOT_FIJOS, fname=fname)
    f_amort = _fila(ws, r'^amortizacion equipamiento', desde=f_fij,
                    hasta=f_tot_fij, fname=fname)
    f_eb = _fila(ws, RX_EBITDA, desde=f_tot_fij, fname=fname)
    f_me = _fila(ws, RX_PCT_EBITDA, desde=f_eb, fname=fname)

    # TOTAL COSTES FIJOS sin la amortización: lo que la hoja llamaba EBITDA era
    # un EBIT. Con ventas 140.000 € y amortización 6.000 €, el R1 midió 26.000 €
    # donde el EBITDA es 32.000 € (−23 %).
    # RT-29/RC-34: un `SUM(B31:B31)` sobre UNA celda invita a creer que el
    # rango se extiende si alguien inserta una fila, y no dice por qué la 30
    # está fuera. §2.3.2 lo escribe como `+B31`.
    def _tramo(desde, hasta):
        return ('SUM(B' + str(desde) + ':B' + str(hasta) + ')'
                if hasta > desde else 'B' + str(desde))
    trozos = []
    if f_amort > f_fij + 1:
        trozos.append(_tramo(f_fij + 1, f_amort - 1))
    if f_amort < f_tot_fij - 1:
        trozos.append(_tramo(f_amort + 1, f_tot_fij - 1))
    # RC-09 · la decisión 13 de §7-bis («sin dato se escribe "", nunca 0») se
    # había aplicado a la columna de porcentajes y NO a la de importes, que es
    # la que se lee: con el libro vaciado, «TOTAL INGRESOS», «EBITDA» y «EBIT»
    # imprimían «0,00 €» con formato de euro, como si fueran resultados. La
    # guarda es `COUNT` (números), nunca `COUNTIF(rango,"<>")` — ver RT-01.
    conteo = ','.join(t.replace('SUM(', '').rstrip(')') if t.startswith('SUM(')
                      else t for t in trozos)
    motor.f(ws, 'B' + str(f_tot_fij),
            '=IF(COUNT(' + conteo + ')=0,"",' + '+'.join(trozos) + ')',
            fmt=FMT_EUR)
    motor.f(ws, 'B' + str(f_tot_ing),
            '=IF(COUNT(B' + str(f_ing_bloque + 1) + ':B' + str(f_tot_ing - 1)
            + ')=0,"",SUM(B' + str(f_ing_bloque + 1) + ':B'
            + str(f_tot_ing - 1) + '))', fmt=FMT_EUR, bold=True)
    motor.f(ws, 'B' + str(f_tot_var),
            '=IF(COUNT(B' + str(f_var_bloque + 1) + ':B' + str(f_tot_var - 1)
            + ')=0,"",SUM(B' + str(f_var_bloque + 1) + ':B'
            + str(f_tot_var - 1) + '))', fmt=FMT_EUR, bold=True)
    motor.f(ws, 'B' + str(f_mb),
            '=IF(OR(B' + str(f_tot_ing) + '="",B' + str(f_tot_var)
            + '=""),"",B' + str(f_tot_ing) + '-B' + str(f_tot_var) + ')',
            fmt=FMT_EUR, bold=True)
    motor.f(ws, 'B' + str(f_eb),
            '=IF(OR(B' + str(f_mb) + '="",B' + str(f_tot_fij) + '=""),"",B'
            + str(f_mb) + '-B' + str(f_tot_fij) + ')', fmt=FMT_EUR, bold=True)
    motor.val(ws, 'A' + str(f_tot_fij),
              'TOTAL COSTES FIJOS (sin amortización)')
    nota(ws, 'D' + str(f_tot_fij),
         'No incluye la amortización de la fila ' + str(f_amort) + ': por eso '
         'la fila de abajo es EBITDA y no EBIT.')
    # 2.3.3 · el margen es un ratio, no un importe (TEC-09) — se CLAVA porque
    # la cabecera de la columna B dice «Importe (€)» y la regla de columna del
    # §1.4 lo devolvería a euros.
    motor.f(ws, 'B' + str(f_me),
            '=IF(OR(B' + str(f_tot_ing) + '="",B' + str(f_tot_ing)
            + '=0,B' + str(f_eb) + '=""),"",B' + str(f_eb) + '/B'
            + str(f_tot_ing) + ')')
    motor.fijar_formato(ws, 'B' + str(f_me), FMT_PCT)

    # filas nuevas: la amortización recuperada y el EBIT
    f_amo2, f_ebit = f_me + 1, f_me + 2
    motor.val(ws, 'A' + str(f_amo2),
              'Amortización (fuera del EBITDA, ver fila ' + str(f_amort) + ')')
    motor.f(ws, 'B' + str(f_amo2),
            '=IF(B' + str(f_amort) + '="","",B' + str(f_amort) + ')',
            fmt=FMT_EUR)
    motor.val(ws, 'A' + str(f_ebit), 'EBIT (resultado de explotación)')
    motor.f(ws, 'B' + str(f_ebit),
            '=IF(B' + str(f_eb) + '="","",B' + str(f_eb) + '-IF(ISNUMBER(B'
            + str(f_amo2) + '),B' + str(f_amo2) + ',0))', fmt=FMT_EUR,
            bold=True)

    # 2.3.3 · la columna «% s/Ventas» deja de decir «0,0 %» con el libro en
    # blanco (§7-bis.13) y C35 deja de dividir un porcentaje entre la
    # facturación (TEC-10): esa fila ya no tiene eco en la columna de %.
    # Criterio: es fila de DATO la que tiene etiqueta y no es un encabezado de
    # bloque. No sirve mirar si B está vacía —en el representante lo está
    # ENTERA, porque el P&L es del cliente— y vaciar por eso la columna de %
    # borraría 24 fórmulas legítimas.
    rx_bloque = re.compile(r'^(ingresos|costes variables|costes fijos)$')
    limpiadas, rehechas = [], 0
    for r, norm, _c in etiquetas(ws, 1, 5, f_ebit):
        if r == f_me or rx_bloque.match(norm):
            if ws['C' + str(r)].value is not None:
                motor.limpiar_rango(ws, 'C' + str(r))
                limpiadas.append('C' + str(r))
            continue
        motor.f(ws, 'C' + str(r),
                '=IF(OR($B$' + str(f_tot_ing) + '="",$B$' + str(f_tot_ing)
                + '=0,B' + str(r) + '=""),"",B' + str(r) + '/$B$'
                + str(f_tot_ing) + ')', fmt=FMT_PCT)
        rehechas += 1
    _semaforo_negativo(ws, ['B'], (f_eb, f_ebit))
    motor.permitir_negativo(ws, 'B' + str(f_eb) + ':B' + str(f_ebit))
    _precargar_pl_mensual(ws, fname, cambios, contenido, f_ebit)
    apunta(cambios, fname, ws,
           'EBITDA sin amortización (B' + str(f_tot_fij) + '=' + '+'.join(trozos)
           + '), EBIT en la fila ' + str(f_ebit) + ', margen en % (TEC-08/09), '
           + str(rehechas) + ' celdas de «% s/Ventas» con "" en vez de 0 y '
           + str(len(limpiadas)) + ' vaciadas ' + str(limpiadas) + ' (TEC-10)')
    f_alq = _fila(ws, r'^alquiler', desde=f_fij, hasta=f_tot_fij, fname=fname)
    f_pcoc = _fila(ws, r'^personal cocina', desde=f_fij, hasta=f_tot_fij,
                   obligatoria=False) or f_alq
    f_psal = _fila(ws, r'^personal sala', desde=f_fij, hasta=f_tot_fij,
                   obligatoria=False) or f_pcoc
    return {'tot_ing': f_tot_ing, 'tot_var': f_tot_var, 'margen': f_mb,
            'tot_fij': f_tot_fij, 'amort': f_amort, 'ebitda': f_eb,
            'ebit': f_ebit, 'alquiler': f_alq, 'personal': (f_pcoc, f_psal)}


def _precargar_pl_mensual(ws, fname, cambios, contenido, ultima):
    """RD-01/RT-02/RC-01 — el libro que §7-bis.7 declara FUENTE ÚNICA se
    entregaba con la columna de entrada VACÍA: «TOTAL INGRESOS 0,00 €»,
    «EBITDA 0,00 €» y «Proyección 3 Años» entera en blanco.

    Los importes salen de `contenido_<pid>/a.py` y son el escenario REALISTA
    de `pl-mensual-escenarios.xlsx`, línea a línea, con su nota de origen: van
    en VERDE y son valores de EJEMPLO (§1.2). Si el módulo de contenido no los
    trae, la hoja se queda como está: aquí no se inventa ninguna cifra.
    """
    conf = ((getattr(contenido, 'PLAN', None) or {}) if contenido else {})
    valores = conf.get('pl_mensual') or {}
    if not valores:
        return 0
    puestas = 0
    for patron, valor in valores.items():
        r = _fila(ws, patron, desde=5, hasta=ultima, obligatoria=False)
        if r is None or ws['B' + str(r)].data_type == 'f':
            continue
        motor.val(ws, 'B' + str(r), valor, fmt=FMT_EUR, verde_=True)
        motor.fijar_formato(ws, 'B' + str(r), FMT_EUR)
        puestas += 1
    for patron, texto in (conf.get('pl_notas') or {}).items():
        r = _fila(ws, patron, desde=5, hasta=ultima, obligatoria=False)
        if r is not None:
            nota(ws, 'D' + str(r), texto)
    if puestas:
        apunta(cambios, fname, ws, str(puestas) + ' líneas del P&L precargadas '
               'en verde con el escenario REALISTA de pl-mensual-escenarios '
               '(RD-01/RT-02/RC-01: el libro ya no abre con 0,00 € en TOTAL '
               'INGRESOS, EBITDA y EBIT)')
    return puestas


# --------------------------------------------------------------------------
# 2.5 NUEVO-01 · el P&L Mensual de los 5 hermanos no tiene NI UN total
# --------------------------------------------------------------------------
def _pl_mensual_hermanos(wb, fname, cambios, contenido):
    ws = wb['P&L Mensual']
    fila_cab = 4
    meses, total = columnas_mes(ws, fila_cab)
    if len(meses) < 12:
        raise VarianteDesconocida(
            fname + ':' + ws.title + ': se esperaban 12 columnas de mes y se '
            'han leído ' + str(len(meses)) + ' en la cabecera de la fila '
            + str(fila_cab))
    f_ing = _fila(ws, RX_BLOQUE_ING, fname=fname)
    f_tot_ing = _fila(ws, RX_TOT_INGRESOS, fname=fname)
    f_var = _fila(ws, RX_BLOQUE_VAR, fname=fname)
    f_fc = _fila(ws, RX_FOOD_COST, desde=f_var, fname=fname)
    f_tot_var = _fila(ws, RX_TOT_VARIABLES, fname=fname)
    f_fij = _fila(ws, RX_BLOQUE_FIJ, fname=fname)
    f_tot_fij = _fila(ws, RX_TOT_FIJOS, fname=fname)
    f_eb = _fila(ws, RX_EBITDA, desde=f_tot_fij, fname=fname)
    f_pct = _fila(ws, RX_PCT_EBITDA, desde=f_eb, obligatoria=False)
    conf = getattr(contenido, 'PLAN', None) if contenido else None
    celda_fc = _celda_food_cost_grid(ws, f_fc, conf, cambios, fname, meses)

    for col in meses + ([total] if total else []):
        if col == total:
            motor.f(ws, col + str(f_tot_ing),
                    '=SUM(' + meses[0] + str(f_tot_ing) + ':' + meses[-1]
                    + str(f_tot_ing) + ')', fmt=FMT_EUR)
            motor.f(ws, col + str(f_fc),
                    '=SUM(' + meses[0] + str(f_fc) + ':' + meses[-1]
                    + str(f_fc) + ')', fmt=FMT_EUR)
        else:
            motor.f(ws, col + str(f_tot_ing),
                    '=SUM(' + col + str(f_ing + 1) + ':' + col
                    + str(f_tot_ing - 1) + ')', fmt=FMT_EUR)
            motor.f(ws, col + str(f_fc),
                    '=' + col + str(f_tot_ing) + '*' + celda_fc, fmt=FMT_EUR)
        motor.f(ws, col + str(f_tot_var),
                '=SUM(' + col + str(f_fc) + ':' + col + str(f_tot_var - 1)
                + ')', fmt=FMT_EUR)
        motor.f(ws, col + str(f_tot_fij),
                '=SUM(' + col + str(f_fij + 1) + ':' + col + str(f_tot_fij - 1)
                + ')', fmt=FMT_EUR)
        motor.f(ws, col + str(f_eb),
                '=' + col + str(f_tot_ing) + '-' + col + str(f_tot_var) + '-'
                + col + str(f_tot_fij), fmt=FMT_EUR, bold=True)
        if f_pct:
            motor.f(ws, col + str(f_pct),
                    '=IF(' + col + str(f_tot_ing) + '=0,"",' + col + str(f_eb)
                    + '/' + col + str(f_tot_ing) + ')', fmt=FMT_PCT)
    todas = meses + ([total] if total else [])
    _semaforo_negativo(ws, todas, [f_eb] + ([f_pct] if f_pct else []))
    apunta(cambios, fname, ws, 'NUEVO-01: totales, food cost, EBITDA y % '
           'EBITDA calculados en las ' + str(len(todas)) + ' columnas '
           '(filas ' + str(f_tot_ing) + '/' + str(f_fc) + '/' + str(f_tot_var)
           + '/' + str(f_tot_fij) + '/' + str(f_eb) + ')')
    return {'tot_ing': f_tot_ing, 'tot_var': f_tot_var, 'tot_fij': f_tot_fij,
            'ebitda': f_eb, 'meses': meses, 'total': total, 'amort': None}


def _celda_food_cost_grid(ws, f_fc, conf, cambios, fname, meses):
    """El food cost del rótulo («Food cost (33%)») a una celda verde ABSOLUTA.

    Se coloca en la primera columna libre a la derecha de la rejilla para no
    pisar ningún mes.
    """
    from openpyxl.utils import get_column_letter as gcl, column_index_from_string
    crudo = ws['A' + str(f_fc)].value or ''
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', crudo)
    pct = float(m.group(1).replace(',', '.')) / 100 if m else None
    if (conf or {}).get('food_cost') is not None:
        pct = conf['food_cost']
    col = gcl(column_index_from_string(meses[-1]) + 2)
    destino = col + str(f_fc)
    motor.val(ws, col + str(f_fc - 1), 'Food cost objetivo (%)')
    if pct is not None and ws[destino].data_type != 'f':
        motor.val(ws, destino, pct, fmt=FMT_PCT, verde_=True)
    motor.fijar_formato(ws, destino, FMT_PCT)
    apunta(cambios, fname, ws, 'food cost ' + repr(pct) + ' extraído del '
           'rótulo ' + repr(crudo[:34]) + ' a la celda verde ' + destino)
    return '$' + col + '$' + str(f_fc)


# --------------------------------------------------------------------------
# 2.3.4 · hoja «Financiación» — cuadro francés con anualidad ALGEBRAICA
# --------------------------------------------------------------------------
def hoja_financiacion(wb, fname, cambios, contenido, subtitulo):
    """DOM-22: hoy no hay una sola línea de préstamo en los 141 ficheros.

    `pycel` **no implementa `PMT`** (SPEC, cabecera): la cuota va como anualidad
    algebraica `importe*i/(1-(1+i)^-n)` con el tipo MENSUAL, que es como se paga
    un préstamo de verdad. El cuadro es anual pero se deriva del mismo préstamo
    mensual con la forma cerrada del capital pendiente
    `P*((1+i)^N-(1+i)^k)/((1+i)^N-1)`, verificada contra la amortización mes a
    mes (100.000 € al 5 % en 60 meses: 43.014,90 € pendientes al cierre del año
    3 por las dos vías). Así la cuota que se lleva al cash flow y los intereses
    que se lleva la proyección **son del mismo préstamo**: un cuadro anual
    calculado con el tipo anual daría 23.097 €/año donde la cuota mensual paga
    22.645 €/año, y las dos cifras del producto se contradirían.

    Dos guardas heredadas del kit plan-financiero, que allí se pagaron caras:
      · **carencia ≥ plazo** → aviso de TEXTO en la columna `Cuota`, dejando
        `Capital inicial`, `Intereses` y `Capital pendiente` NUMÉRICAS: de los
        intereses vive el P&L y del capital pendiente el encadenado del año
        siguiente, así que un texto ahí propagaría `#¡VALOR!`.
      · **pasado el vencimiento** el cuadro se apaga a `0` numérico, o el
        capital pendiente se vuelve negativo y el préstamo «cobra» al banco.
    """
    ws, nueva = hoja(wb, HOJA_FIN, tras='P&L Mensual')
    conf = ((getattr(contenido, 'PLAN', None) or {}).get('financiacion')
            if contenido else None) or {}
    cabecera(ws, 'Financiación del proyecto', subtitulo,
             ['Año', 'Capital inicial (€)', 'Cuota del año (€)',
              'Intereses (€)', 'Amortización del principal (€)',
              'Capital pendiente (€)'], fila=17, ancho_a=48)
    motor.val(ws, 'A4', 'PARÁMETROS DEL PRÉSTAMO', bold=True)
    campos = (
        (5, 'Importe del préstamo (€)', conf.get('importe'), FMT_EUR),
        (6, 'Plazo (años)', conf.get('plazo'), FMT_ENT),
        (7, 'Tipo de interés nominal anual (%)', conf.get('tipo'), FMT_PCT),
        (8, 'Carencia (años, sólo intereses)', conf.get('carencia'), FMT_ENT))
    for fila, etiqueta, valor, fmt in campos:
        motor.val(ws, 'A' + str(fila), etiqueta)
        motor.val(ws, 'B' + str(fila), valor, fmt=fmt, verde_=True)
        motor.fijar_formato(ws, 'B' + str(fila), fmt)
    motor.val(ws, 'A10', 'Tipo mensual (%)')
    motor.f(ws, 'B10', '=B7/12')
    motor.fijar_formato(ws, 'B10', '0.000%')
    motor.val(ws, 'A11', 'Cuotas de amortización (meses)')
    motor.f(ws, 'B11', '=(B6-B8)*12', fmt=FMT_ENT)
    motor.val(ws, 'A12',
              'Cuota MENSUAL tras la carencia (€) — la que va al cash flow')
    motor.f(ws, 'B12',
            '=IF(B5=0,"",IF(B11<=0,"",IFERROR(B5*B10/(1-(1+B10)^-B11),"")))',
            fmt=FMT_EUR, bold=True)
    motor.val(ws, 'A13', 'Cuota mensual DURANTE la carencia (€) — sólo '
                         'intereses')
    motor.f(ws, 'B13', '=IF(B5=0,"",IF(B8<=0,"",B5*B10))', fmt=FMT_EUR)
    motor.val(ws, 'A14', 'Meses de carencia')
    motor.f(ws, 'B14', '=B8*12', fmt=FMT_ENT)
    nota(ws, 'A15',
         'La cuota se calcula como anualidad: importe × i / (1 − (1+i)^−n), con '
         'i mensual y n el número de cuotas. Durante la carencia sólo se pagan '
         'intereses y el capital pendiente no baja.')
    nota(ws, 'A16',
         'El cuadro cubre ' + str(ANOS_CUADRO) + ' años; si tu plazo es mayor, '
         'copia la última fila hacia abajo. La cuota mensual de B12 es la que '
         'se repite en cash-flow-break-even.xlsx (no hay enlace entre libros: '
         'un .xlsx movido de carpeta daría #REF!).')

    fila0 = 18
    for i in range(ANOS_CUADRO):
        r = fila0 + i
        motor.val(ws, 'A' + str(r), i + 1, fmt=FMT_ENT)
        if i == 0:
            motor.f(ws, 'B' + str(r), '=$B$5', fmt=FMT_EUR)
        else:
            motor.f(ws, 'B' + str(r), '=F' + str(r - 1), fmt=FMT_EUR)
        # capital pendiente por la forma cerrada del préstamo MENSUAL
        motor.f(ws, 'F' + str(r),
                '=IF($B$8>=$B$6,B' + str(r) + ',IF(A' + str(r) + '>=$B$6,0,'
                'IFERROR($B$5*((1+$B$10)^$B$11-(1+$B$10)^MIN($B$11,MAX(0,12*A'
                + str(r) + '-$B$14)))/((1+$B$10)^$B$11-1),B' + str(r) + ')))',
                fmt=FMT_EUR)
        motor.f(ws, 'C' + str(r),
                '=IF($B$8>=$B$6,"' + AVISO_CARENCIA + '",'
                'IF(A' + str(r) + '>$B$6,0,'
                'IF(A' + str(r) + '<=$B$8,$B$5*$B$7,$B$12*12)))', fmt=FMT_EUR)
        motor.f(ws, 'E' + str(r),
                '=IF(ISNUMBER(C' + str(r) + '),B' + str(r) + '-F' + str(r)
                + ',0)', fmt=FMT_EUR)
        motor.f(ws, 'D' + str(r),
                '=IF(ISNUMBER(C' + str(r) + '),C' + str(r) + '-E' + str(r)
                + ',B' + str(r) + '*$B$7)', fmt=FMT_EUR)
    ws.column_dimensions['A'].width = 48.0
    motor.regla_expresion(ws, 'F' + str(fila0) + ':F'
                          + str(fila0 + ANOS_CUADRO - 1),
                          '=AND(ISNUMBER($F' + str(fila0) + '),$F' + str(fila0)
                          + '<0)')
    apunta(cambios, fname, ws, ('hoja CREADA' if nueva else 'hoja rehecha')
           + ': parámetros, cuota mensual (la del cash flow) y cuadro de '
           + str(ANOS_CUADRO) + ' años derivado del MISMO préstamo mensual, '
             'con las dos guardas (DOM-22, §2.3.4)')
    return {'primer_ano': fila0, 'cuota_mensual': 'B12'}


# --------------------------------------------------------------------------
# 2.3.1 · hoja «Proyección 3 Años» (TEC-07, DOM-06, COM-07)
# --------------------------------------------------------------------------
def hoja_proyeccion(wb, fname, cambios, contenido, subtitulo, pl, fin):
    """La pestaña que `Instrucciones!A9` anuncia y que no existe en ninguna
    de las 7 guías.

    Año 1 va **por referencia** al P&L mensual (nunca constantes) y toda la
    cadena está guardada por `$B$<ingresos>=""`: con el libro en blanco la hoja
    devuelve `""`, no una proyección de ceros (§7-bis.13).
    """
    ws, nueva = hoja(wb, HOJA_PROY, tras='P&L Mensual')
    conf = ((getattr(contenido, 'PLAN', None) or {}).get('proyeccion')
            if contenido else None) or {}
    cabecera(ws, 'Proyección a 3 años', subtitulo,
             ['Concepto', 'Año 1', 'Año 2', 'Año 3', 'Notas'], fila=4,
             ancho_a=46)
    ws.column_dimensions['E'].width = 52.0
    # El Año 1 sale del P&L **anualizado**, y anualizar no es siempre «×12»:
    # el representante tiene UN mes (B) y los hermanos una rejilla de 12 meses
    # con columna de TOTAL. Multiplicar por 12 la columna de enero de un
    # hermano daría el año sólo si los doce meses fuesen idénticos — que es
    # justo lo que el cliente va a dejar de hacer en cuanto meta estacionalidad.
    col_total = pl.get('total')

    def _anual(fila):
        if not fila:
            return None
        if col_total:
            return "'P&L Mensual'!" + col_total + str(fila)
        return "'P&L Mensual'!B" + str(fila) + '*12'

    def _mensual(fila):
        """El mes de CRUCERO. Es lo que la rampa del año 1 multiplica, y no es
        siempre «la celda de enero»: en la rejilla de 12 meses de los hermanos
        el crucero es la columna TOTAL dividida entre 12."""
        if not fila:
            return None
        if col_total:
            return "('P&L Mensual'!" + col_total + str(fila) + '/12)'
        return "'P&L Mensual'!B" + str(fila)

    pl_ing_celda = ("'P&L Mensual'!" + (col_total or 'B') + str(pl['tot_ing']))
    pl_fij = _anual(pl['tot_fij'])
    pl_amo = _anual(pl.get('amort'))
    mens_ing = _mensual(pl['tot_ing'])
    mens_var = _mensual(pl['tot_var'])

    motor.val(ws, 'A5', 'PARÁMETROS', bold=True)
    motor.val(ws, 'A6', 'Crecimiento de ventas (%)')
    motor.val(ws, 'A7', 'Inflación de costes (%)')
    motor.val(ws, 'A8', 'Tipo del Impuesto de Sociedades (%)')
    motor.val(ws, 'A9', ET_RAMPA_MESES)
    motor.val(ws, 'A10', ET_RAMPA_PRIMERO)
    nota(ws, 'E6', 'Sobre el año anterior. El Año 1 es la base: sale del P&L '
                   'mensual de este mismo libro.')
    nota(ws, 'E8', 'Tipo general 25 %. Una entidad de NUEVA CREACIÓN tributa '
                   'al 15 % en el primer ejercicio con base positiva y en el '
                   'siguiente (art. 29.1 LIS).')
    for col in ('C', 'D'):
        motor.val(ws, col + '6', (conf.get('crecimiento') or {}).get(col),
                  fmt=FMT_PCT, verde_=True)
        motor.val(ws, col + '7', (conf.get('inflacion') or {}).get(col),
                  fmt=FMT_PCT, verde_=True)
        motor.permitir_negativo(ws, col + '6:' + col + '7')
        # RT-14 · una DV `decimal >= -1.000.000.000.000` no valida nada, y
        # estas dos celdas gobiernan la proyección entera.
        motor.dv_propia(ws, col + '6:' + col + '7', -0.5, 1.0, 'Porcentaje',
                        'Escribe un porcentaje entre -0,5 y 1 (0,08 = 8 %). '
                        'Admite negativo: una caída de ventas es -0,10.',
                        'Se escribe en tanto por uno: 0,08 = 8 %. Admite '
                        'negativo para una caída de ventas.')
        motor.semaforo_isnumber(ws, col + '6', col + '6', operador='>',
                                umbral='0.3', bg=motor.CF_AMBAR_BG,
                                fg=motor.CF_AMBAR_FG)
    # El tipo del IS NO se deja vacío aunque falte el módulo de contenido: una
    # celda de impuesto en blanco hace que la línea «Impuesto de Sociedades»
    # valga 0 sin decirlo, y el resultado neto sale inflado. El 25 % viene de
    # la SPEC §2.3.1, no de una estimación del sector.
    tipo_is = conf.get('impuesto_sociedades', 0.25)
    for col in ('B', 'C', 'D'):
        motor.val(ws, col + '8', tipo_is, fmt=FMT_PCT, verde_=True)

    # ---- B-01 · los dos parámetros de la RAMPA del año 1 ------------------
    meses_rampa, pct_primero, suma_rampa = _rampa_ano_1(contenido, conf)
    motor.val(ws, 'B9', meses_rampa, fmt=FMT_ENT, verde_=True)
    motor.fijar_formato(ws, 'B9', FMT_ENT)
    motor.dv_propia(ws, 'B9', 0, 12, 'Meses de rampa',
                    'Un número entero entre 0 y 12. 0 = abres ya al ritmo de '
                    'crucero (el año 1 sería el mes tipo × 12).',
                    'Meses que tardas en llegar al mes de crucero. Con 0 el '
                    'año 1 es el P&L mensual × 12.')
    motor.val(ws, 'B10', pct_primero, fmt=FMT_PCT, verde_=True)
    motor.fijar_formato(ws, 'B10', FMT_PCT)
    motor.dv_propia(ws, 'B10', 0, 1, 'Porcentaje del crucero',
                    'Se escribe en tanto por uno: 0,60 = 60 % del mes de '
                    'crucero.',
                    'Qué parte del mes de crucero facturas el PRIMER mes. '
                    'Entre 0 y 1 (0,60 = 60 %).')

    F_ING = 13
    filas = ((F_ING, 'Ingresos (€)'),
             (F_ING + 1, 'Costes variables (€)'),
             (F_ING + 2, 'Margen bruto (€)'),
             (F_ING + 3, 'Costes fijos sin amortización (€)'),
             (F_ING + 4, 'EBITDA (€)'), (F_ING + 5, 'Amortización (€)'),
             (F_ING + 6, 'EBIT — resultado de explotación (€)'),
             (F_ING + 7, 'Gastos financieros (€)'),
             (F_ING + 8, 'BAI — beneficio antes de impuestos (€)'),
             (F_ING + 9, 'Impuesto de Sociedades (€)'),
             (F_ING + 10, 'Resultado neto (€)'),
             (F_ING + 11, 'Margen EBITDA (%)'))
    motor.val(ws, 'A' + str(F_ING - 1), 'CUENTA DE RESULTADOS PREVISIONAL',
              bold=True)
    for r, etiqueta in filas:
        motor.val(ws, 'A' + str(r), etiqueta)
    guarda = '=IF($B$' + str(F_ING) + '="","",'

    # RC-09 · el P&L devuelve ahora `""` cuando no hay dato, así que la guarda
    # de esta hoja no puede seguir siendo `=0`: `""=0` es FALSO en Excel y la
    # proyección intentaría multiplicar una cadena por 12.
    # El AÑO 1 no es el mes de crucero × 12: la casa abre y factura menos hasta
    # llegar a él. `S_RAMPA` es la suma de los 12 factores mensuales de una
    # rampa LINEAL que arranca en `$B$10` y llega a 1 en el mes `$B$9`
    # (después, crucero). Con `$B$9 = 0` (o 1, o sin `$B$10`) vale 12 y la
    # proyección vuelve exactamente al `P&L × 12` de antes [B-01].
    S_RAMPA = ('IF(OR($B$9="",$B$9<=1,$B$10=""),12,'
               '$B$9*(1+$B$10)/2+(12-$B$9))')
    r_ing, r_var = F_ING, F_ING + 1
    r_mb, r_fij = F_ING + 2, F_ING + 3
    r_eb, r_amo = F_ING + 4, F_ING + 5
    r_ebit, r_gf = F_ING + 6, F_ING + 7
    r_bai, r_is = F_ING + 8, F_ING + 9
    r_neto, r_meb = F_ING + 10, F_ING + 11
    motor.f(ws, 'B' + str(r_ing),
            '=IF(OR(' + pl_ing_celda + '="",' + pl_ing_celda + '=0),"",'
            + mens_ing + '*' + S_RAMPA + ')', fmt=FMT_EUR)
    # Los costes VARIABLES siguen a las ventas, así que llevan la MISMA rampa
    # (es lo que hace el cash flow de este mismo pack); los fijos y la
    # amortización, no: son fijos.
    motor.f(ws, 'B' + str(r_var),
            motor.iferror(guarda[1:] + mens_var + '*' + S_RAMPA + ')'),
            fmt=FMT_EUR)
    motor.f(ws, 'B' + str(r_fij), motor.iferror(guarda[1:] + pl_fij + ')'),
            fmt=FMT_EUR)
    motor.f(ws, 'B' + str(r_amo),
            motor.iferror(guarda[1:] + (pl_amo + ')' if pl_amo else '0)')),
            fmt=FMT_EUR)
    for i, col in enumerate(('C', 'D')):
        ant = 'BC'[i]
        # Año 2 crece sobre el RITMO DE CRUCERO anualizado, no sobre el año 1:
        # en el año 2 ya no hay rampa, y encadenarlo al año 1 dejaría la
        # proyección facturando MENOS que su propio ritmo de crucero — el
        # primer disparate que caza un analista de riesgos. Con esto, los años
        # 2 y 3 valen exactamente lo mismo que antes de B-01.
        base_ing = ('(' + mens_ing + '*12)') if i == 0 else (ant + str(r_ing))
        base_var = ('(' + mens_var + '*12)') if i == 0 else (ant + str(r_var))
        motor.f(ws, col + str(r_ing),
                guarda + base_ing + '*(1+' + col + '6))', fmt=FMT_EUR)
        motor.f(ws, col + str(r_var),
                guarda + base_var + '*(1+' + col + '6)*(1+' + col + '7))',
                fmt=FMT_EUR)
        motor.f(ws, col + str(r_fij),
                guarda + ant + str(r_fij) + '*(1+' + col + '7))', fmt=FMT_EUR)
        motor.f(ws, col + str(r_amo), guarda + ant + str(r_amo) + ')',
                fmt=FMT_EUR)
    for i, col in enumerate(('B', 'C', 'D')):
        f_fin = fin['primer_ano'] + i
        motor.f(ws, col + str(r_mb),
                guarda + col + str(r_ing) + '-' + col + str(r_var) + ')',
                fmt=FMT_EUR)
        motor.f(ws, col + str(r_eb),
                guarda + col + str(r_mb) + '-' + col + str(r_fij) + ')',
                fmt=FMT_EUR, bold=True)
        motor.f(ws, col + str(r_ebit),
                guarda + col + str(r_eb) + '-' + col + str(r_amo) + ')',
                fmt=FMT_EUR)
        motor.f(ws, col + str(r_gf),
                guarda + "IF(ISNUMBER('" + HOJA_FIN + "'!D" + str(f_fin)
                + "),'" + HOJA_FIN + "'!D" + str(f_fin) + ',0))', fmt=FMT_EUR)
        motor.f(ws, col + str(r_bai),
                guarda + col + str(r_ebit) + '-' + col + str(r_gf) + ')',
                fmt=FMT_EUR)
        motor.f(ws, col + str(r_is),
                guarda + 'IF(' + col + str(r_bai) + '<=0,0,' + col
                + str(r_bai) + '*' + col + '8))', fmt=FMT_EUR)
        motor.f(ws, col + str(r_neto),
                guarda + col + str(r_bai) + '-' + col + str(r_is) + ')',
                fmt=FMT_EUR, bold=True)
        motor.f(ws, col + str(r_meb),
                guarda + 'IFERROR(' + col + str(r_eb) + '/' + col + str(r_ing)
                + ',""))', fmt=FMT_PCT)
    nota(ws, 'E9',
         'Cuántos meses tardas en llegar al ritmo de crucero. Escribe 0 y el '
         'Año 1 vuelve a ser el mes tipo × 12. Con ' + str(meses_rampa)
         + ' y ' + str(round(pct_primero * 100)) + ' %, el Año 1 factura lo '
         'mismo que cash-flow-break-even.xlsx de este mismo pack.')
    nota(ws, 'E10',
         'Qué parte del mes de crucero facturas el PRIMER mes. De ahí sube en '
         'línea recta hasta el 100 % en el mes de crucero.')
    nota(ws, 'E' + str(r_ing),
         'Año 1 = ' + ('columna TOTAL del P&L Mensual' if col_total
                       else 'P&L Mensual') + ' ÷ 12 × la suma de los doce '
         'factores de la rampa de arriba (con 0 meses de rampa, × 12). El '
         'cash flow de este mismo pack reparte la rampa mes a mes; el TOTAL '
         'del año coincide al céntimo. Los años 2 y 3 parten del ritmo de '
         'CRUCERO anualizado, no del año 1, porque en ellos ya no hay rampa. '
         'Si cambias el P&L, cambia la proyección entera: no hay ni una '
         'constante aquí.')
    nota(ws, 'E' + str(r_var),
         'Los costes variables llevan la misma rampa que las ventas; los '
         'fijos y la amortización, no.')
    nota(ws, 'E' + str(r_gf),
         'Intereses del año, tomados del cuadro francés de la hoja '
         '«' + HOJA_FIN + '».')
    _semaforo_negativo(ws, ['B', 'C', 'D'], (r_eb, r_neto, r_meb))
    apunta(cambios, fname, ws, ('hoja CREADA' if nueva else 'hoja rehecha')
           + ': Año 1 por referencia al P&L, EBITDA→EBIT→BAI→IS→resultado '
             'neto, todo guardado con "" (TEC-07, DOM-06, COM-07)')


# --------------------------------------------------------------------------
# 2.3.5 · fondo de maniobra dimensionado  ·  2.3.6 · fuente única de CAPEX
# --------------------------------------------------------------------------
def _precargar_inversion(ws, fname, cambios, contenido, col_pres, col_etq,
                         fila_cab):
    """RD-01/RT-02/RC-01/RC-05 — las 22 partidas de CAPEX se entregaban VACÍAS.

    Los importes vienen de `contenido_<pid>/a.py` y CUADRAN con lo que los
    checklists de este mismo pack ya tasan (equipamiento 164.718,40 €, sala
    108.200 €, vajilla 30.230 €, marketing 23.800 €, legal 23.960 €), que es lo
    que RC-05 pedía reconciliar. Van en VERDE y son ejemplo (§1.2).
    """
    valores = (((getattr(contenido, 'PLAN', None) or {}).get('inversion')
                if contenido else None) or {})
    if not valores:
        return 0
    puestas = 0
    for patron, valor in valores.items():
        r = _fila(ws, patron, col=col_etq, desde=fila_cab + 1,
                  obligatoria=False)
        if r is None or ws[col_pres + str(r)].data_type == 'f':
            continue
        motor.val(ws, col_pres + str(r), valor, fmt=FMT_EUR, verde_=True)
        motor.fijar_formato(ws, col_pres + str(r), FMT_EUR)
        puestas += 1
    if puestas:
        apunta(cambios, fname, ws, str(puestas) + ' partidas de CAPEX '
               'precargadas en verde, cuadradas con los totales que ya tasan '
               'los checklists del propio pack (RD-01, RC-05)')
    return puestas


def fondo_de_maniobra(wb, fname, cambios, contenido, pl, hoja_inv, col_pres,
                      col_etq=1, fila_cab=None):
    """DOM-01/COM-30 + RD-08/RD-20: el fondo de maniobra deja de ser un importe
    tecleado, aparece la PREAPERTURA y el libro dice cuánto dinero hace falta.

    `= (coste fijo mensual sin amortización + coste variable del mes) × meses`,
    con los meses en celda verde y **mínimo 6**, que es lo que pide el propio
    consejo del capítulo 4 («un gastronómico tarda 6-12 meses en alcanzar
    velocidad de crucero»). El rótulo de la partida NO se toca (§2.3.5).

    RD-08: el Gantt de este mismo pack firma el arrendamiento en el mes 3 y
    contrata la brigada en el 12 para abrir en el 18, y no había ni una partida
    de renta ni de nóminas de preapertura en los 18 xlsx. Es la partida que más
    aperturas mata. Se calcula aquí, con los meses en verde y el importe leído
    del P&L, y entra en la NECESIDAD TOTAL DE FINANCIACIÓN, que es la cifra que
    la hoja «Financiación» tiene que cuadrar (RD-20).

    ⚠️ El fallo que esto corrige NO era de diseño sino de LOCALIZACIÓN: la
    función existía y `_fila()` la buscaba en la columna A, donde «Inversión»
    sólo tiene el número de orden. Los conceptos están en la B. Devolvía `None`
    y se salía sin escribir nada, en silencio (RT-03, RC-03).
    """
    ws = wb[hoja_inv]
    f_fondo = _fila(ws, r'^fondo de maniobra', col=col_etq, obligatoria=False)
    if f_fondo is None:
        return None
    conf = ((getattr(contenido, 'PLAN', None) or {}) if contenido else {})
    meses = (conf.get('fondo_maniobra') or {}).get('meses', 6)
    prea = conf.get('preapertura') or {}
    # La preapertura necesita las filas de alquiler y personal del P&L. Los
    # 5 hermanos tienen otra rejilla y `_pl_mensual_hermanos` no las devuelve:
    # allí el bloque se queda en el fondo de maniobra, sin inventar filas.
    con_prea = bool(pl.get('alquiler') and pl.get('personal'))
    # marca propia y DISTINTA del rótulo de la partida: `^fondo de maniobra`
    # encontraría antes la fila de la tabla y el bloque se escribiría encima.
    base = base_bloque(ws, r'^preapertura y fondo de maniobra')
    r = base
    motor.val(ws, 'A' + str(r), 'PREAPERTURA Y FONDO DE MANIOBRA — lo calcula '
              'el libro (la fila ' + str(f_fondo) + ' de la tabla sale de '
              'aquí)', bold=True)
    r += 1
    motor.val(ws, 'A' + str(r), 'Coste mensual de estructura (fijos sin '
              'amortización + variables) (€)')
    motor.f(ws, col_pres + str(r),
            "=IF(OR('P&L Mensual'!B" + str(pl['tot_fij']) + "=\"\",'P&L "
            "Mensual'!B" + str(pl['tot_var']) + "=\"\"),\"\",'P&L Mensual'!B"
            + str(pl['tot_fij']) + "+'P&L Mensual'!B" + str(pl['tot_var'])
            + ')', fmt=FMT_EUR)
    f_estr = r

    f_prea = None
    if con_prea:
        r += 1
        motor.val(ws, 'A' + str(r), 'Alquiler mensual (€)')
        motor.f(ws, col_pres + str(r),
                "=IF('P&L Mensual'!B" + str(pl['alquiler']) + "=\"\",\"\","
                "'P&L Mensual'!B" + str(pl['alquiler']) + ')', fmt=FMT_EUR)
        f_alq = r
        r += 1
        motor.val(ws, 'A' + str(r), 'Coste mensual de la brigada (€)')
        motor.f(ws, col_pres + str(r),
                "=IF(COUNT('P&L Mensual'!B" + str(pl['personal'][0])
                + ",'P&L Mensual'!B" + str(pl['personal'][1]) + ")=0,\"\","
                "SUM('P&L Mensual'!B" + str(pl['personal'][0])
                + ",'P&L Mensual'!B" + str(pl['personal'][1]) + '))',
                fmt=FMT_EUR)
        f_bri = r
        r += 1
        motor.val(ws, 'A' + str(r), 'Meses de renta ANTES de abrir (fuera de '
                  'la carencia que negocies)')
        motor.val(ws, col_pres + str(r), prea.get('meses_renta', 6),
                  fmt=FMT_ENT, verde_=True)
        motor.fijar_formato(ws, col_pres + str(r), FMT_ENT)
        f_mren = r
        r += 1
        motor.val(ws, 'A' + str(r),
                  'Meses de nómina de la brigada ANTES de abrir')
        motor.val(ws, col_pres + str(r), prea.get('meses_nomina', 2),
                  fmt=FMT_ENT, verde_=True)
        motor.fijar_formato(ws, col_pres + str(r), FMT_ENT)
        f_mnom = r
        r += 1
        motor.val(ws, 'A' + str(r), 'Rentas y suministros de preapertura (€)')
        motor.f(ws, col_pres + str(r),
                '=IF(OR(' + col_pres + str(f_alq) + '="",' + col_pres
                + str(f_mren) + '=""),"",' + col_pres + str(f_alq) + '*'
                + col_pres + str(f_mren) + ')', fmt=FMT_EUR)
        f_pren = r
        r += 1
        motor.val(ws, 'A' + str(r), 'Nóminas de preapertura (€)')
        motor.f(ws, col_pres + str(r),
                '=IF(OR(' + col_pres + str(f_bri) + '="",' + col_pres
                + str(f_mnom) + '=""),"",' + col_pres + str(f_bri) + '*'
                + col_pres + str(f_mnom) + ')', fmt=FMT_EUR)
        f_pnom = r
        r += 1
        motor.val(ws, 'A' + str(r), 'TOTAL COSTES DE PREAPERTURA (€)',
                  bold=True)
        motor.f(ws, col_pres + str(r),
                '=IF(COUNT(' + col_pres + str(f_pren) + ':' + col_pres
                + str(f_pnom) + ')=0,"",SUM(' + col_pres + str(f_pren) + ':'
                + col_pres + str(f_pnom) + '))', fmt=FMT_EUR, bold=True)
        f_prea = r
        r += 1
        nota(ws, 'A' + str(r),
             'El cronograma de este mismo pack firma el arrendamiento en el '
             'mes 3 y contrata la brigada en el 12 para abrir en el 18: hay '
             'renta y hay nóminas antes de facturar un euro, y el checklist '
             'legal lo dice («son 6-10 meses de renta antes de facturar»). '
             'Esta partida no está entre los conceptos de la tabla de arriba: '
             'se suma aparte, en la necesidad total.')
    r += 2

    # ---- fondo de maniobra (§2.3.5) --------------------------------------
    motor.val(ws, 'A' + str(r), 'Meses de colchón (mínimo 6)')
    motor.val(ws, col_pres + str(r), meses, fmt=FMT_ENT, verde_=True)
    motor.fijar_formato(ws, col_pres + str(r), FMT_ENT)
    f_meses = r
    r += 1
    motor.val(ws, 'A' + str(r), 'Fondo de maniobra necesario (€)')
    motor.f(ws, col_pres + str(r),
            '=IF(OR(' + col_pres + str(f_estr) + '="",' + col_pres
            + str(f_meses) + '=""),"",' + col_pres + str(f_estr) + '*'
            + col_pres + str(f_meses) + ')', fmt=FMT_EUR, bold=True)
    f_fmn = r
    r += 1
    nota(ws, 'A' + str(r),
         'Los 60.000 / 120.000 / 200.000 € que trae calculadora-capex.xlsx no '
         'cubren ni dos meses de la estructura que describe el propio libro. '
         'Aquí se dimensiona con TUS costes: rellena el P&L Mensual y esta '
         'fila se calcula sola. El coste de personal sale de '
         'plantilla-turnos-brigada.xlsx (§7-bis.7).')
    r += 2

    # ---- necesidad total (RD-20) -----------------------------------------
    f_total_tabla = _fila(ws, r'^total', col=col_etq,
                          desde=(fila_cab or 4) + 1, hasta=base - 1,
                          obligatoria=False)
    f_neces = None
    if f_total_tabla:
        motor.val(ws, 'A' + str(r), 'NECESIDAD TOTAL DE FINANCIACIÓN (€) — '
                  'CAPEX de la tabla + preapertura', bold=True)
        suma = col_pres + str(f_total_tabla)
        if f_prea:
            suma += ('+IF(ISNUMBER(' + col_pres + str(f_prea) + '),'
                     + col_pres + str(f_prea) + ',0)')
        motor.f(ws, col_pres + str(r),
                '=IF(OR(' + col_pres + str(f_total_tabla) + '="",' + col_pres
                + str(f_total_tabla) + '=0),"",' + suma + ')',
                fmt=FMT_EUR, bold=True)
        f_neces = r
        r += 1
        nota(ws, 'A' + str(r),
             'Esta es la cifra que tiene que cuadrar la hoja «' + HOJA_FIN
             + '»: fondos propios + préstamo + otras fuentes. El fondo de '
             'maniobra ya está dentro, en la fila ' + str(f_fondo) + ' de la '
             'tabla.')

    # la partida de la tabla toma el valor calculado
    motor.f(ws, col_pres + str(f_fondo),
            '=IF(' + col_pres + str(f_fmn) + '="","",' + col_pres + str(f_fmn)
            + ')', fmt=FMT_EUR)
    motor.regla_expresion(ws, col_pres + str(f_meses),
                          '=AND(ISNUMBER(' + col_pres + str(f_meses) + '),'
                          + col_pres + str(f_meses) + '<6)')
    apunta(cambios, fname, ws, 'fondo de maniobra CALCULADO en ' + col_pres
           + str(f_fondo) + ' = estructura mensual × meses (verde en '
           + col_pres + str(f_meses) + ', mínimo 6)'
           + (', preapertura en ' + col_pres + str(f_prea) if f_prea else '')
           + (', necesidad total en ' + col_pres + str(f_neces)
              if f_neces else '')
           + ' — DOM-01, COM-30, RD-08, RD-20')
    return {'base': base, 'necesidad': f_neces, 'preapertura': f_prea,
            'fondo': f_fmn}


def correspondencia_capex(wb, fname, cambios, contenido, hoja_inv, fila_cab,
                          col_etq=1):
    """§2.3.6 (TEC-26, COM-32) — los dos CAPEX se reconcilian SIN fusionar.

    `calculadora-capex.xlsx` queda como **hoja de rangos de mercado** y
    `plan-financiero!'Inversión'` como **«Mi CAPEX»**. La correspondencia se
    escribe en una columna nueva de la propia hoja, no con `externalLink`: un
    `.xlsx` movido de carpeta daría `#REF!` al cliente (§1.13, §7.1).

    RT-04/RC-04: la columna y la frase de «Instrucciones» que la anuncia ya
    estaban; las 22 celdas seguían VACÍAS porque `_fila()` buscaba el concepto
    en la columna A (la del número de orden) en vez de en la B.
    """
    from openpyxl.utils import get_column_letter as gcl
    mapa = ((getattr(contenido, 'PLAN', None) or {}).get('capex_map')
            if contenido else None)
    if not mapa:
        return 0
    ws = wb[hoja_inv]
    titulo = 'Categoría equivalente en calculadora-capex.xlsx'
    # la columna se busca por su CABECERA: `ws.max_column + 1` añadiría una
    # columna nueva en cada pasada (misma trampa que la de `ws.max_row + 2`).
    indice = None
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(row=fila_cab, column=c).value) == _norm(titulo):
            indice = c
            break
    if indice is None:
        indice = ws.max_column + 1
    col = gcl(indice)
    cel = ws.cell(row=fila_cab, column=indice, value=titulo)
    cel.font = Font(bold=True, color='FFFFFF', size=10)
    cel.fill = motor.PatternFill('solid', fgColor='2D2D2D')
    cel.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.column_dimensions[col].width = 34.0
    puestas = 0
    for patron, categoria in mapa.items():
        r = _fila(ws, patron, col=col_etq, desde=fila_cab + 1,
                  obligatoria=False)
        if r is None:
            continue
        motor.val(ws, col + str(r), categoria)
        puestas += 1
    apunta(cambios, fname, ws, 'columna ' + col + ' con la correspondencia de '
           + str(puestas) + ' conceptos hacia las categorías de '
           'calculadora-capex.xlsx (§2.3.6, TEC-26/COM-32)')
    return puestas


def instruccion(wb, texto, rx=None):
    """Línea en la hoja `Instrucciones`, sin duplicar en la 2.ª pasada."""
    if 'Instrucciones' not in wb.sheetnames:
        return None
    return motor.linea_instrucciones(wb['Instrucciones'], texto, rx)


RX_INSTR_FUENTE = re.compile(r'^Fuente de cada cifra:')
RX_INSTR_PROY = re.compile(r'^La pesta.a «Proyecci')
RX_INSTR_FIN = re.compile(r'^La pesta.a «Financiaci')
#: Aviso del crítico: `'Inversión'!D27` es un TOTAL que evalúa a `""` porque
#: `D5:D26` están vacías —correcto por §1 («sin dato» se escribe `""`, nunca
#: 0)—, pero una fila TOTAL en blanco al lado de otra con 1.668.340,88 € se lee
#: como un error. La hoja lo dice ahora en Instrucciones.
RX_INSTR_REAL = re.compile(r'^La columna «Real')


def _columna_concepto(ws, fila_cab):
    """Índice (1-based) de la columna cuya cabecera es el CONCEPTO de la fila.

    En «Inversión» la columna A es el `#` y el concepto está en la B: leerlo de
    la cabecera evita el `col=1` por defecto de `_fila()`, que es lo que dejó
    §2.3.5 y §2.3.6 sin escribir (RT-03, RT-04, RC-03, RC-04).
    """
    rx = re.compile(r'^(concepto|categoria|partida|descripcion)')
    for c in range(1, ws.max_column + 1):
        if rx.match(_norm(ws.cell(row=fila_cab, column=c).value)):
            return c
    return 1


def plan_de_financiacion(wb, fname, cambios, contenido, inv):
    """RD-20 — «préstamo + fondos propios = necesidad» es lo PRIMERO que cuadra
    un analista de riesgos, y la hoja no lo planteaba: traía 300.000 € de
    préstamo precargados contra un CAPEX de 0 € y ni una fila de aportación.
    """
    if not inv or HOJA_FIN not in wb.sheetnames:
        return None
    ws = wb[HOJA_FIN]
    conf = (((getattr(contenido, 'PLAN', None) or {}).get('financiacion')
             if contenido else None) or {})
    f_imp = _fila(ws, r'^importe del prestamo', obligatoria=False)
    if f_imp is None:
        return None
    base = base_bloque(ws, r'^plan de financiacion')
    r = base
    motor.val(ws, 'A' + str(r), 'PLAN DE FINANCIACIÓN — tiene que cuadrar con '
              'la hoja «Inversión»', bold=True)
    r += 1
    motor.val(ws, 'A' + str(r), 'Necesidad total de financiación (€)')
    motor.f(ws, 'B' + str(r),
            "=IF('Inversión'!C" + str(inv['necesidad']) + '="","",'
            + "'Inversión'!C" + str(inv['necesidad']) + ')', fmt=FMT_EUR)
    f_nec = r
    nota(ws, 'C' + str(r), 'Sale de la hoja «Inversión» de este mismo libro: '
                           'los 22 conceptos de CAPEX (con el fondo de '
                           'maniobra ya calculado) más los costes de '
                           'preapertura.')
    r += 1
    motor.val(ws, 'A' + str(r), 'Fondos propios / aportación de socios (€)')
    motor.val(ws, 'B' + str(r), conf.get('fondos_propios'), fmt=FMT_EUR,
              verde_=True)
    motor.fijar_formato(ws, 'B' + str(r), FMT_EUR)
    f_fp = r
    r += 1
    motor.val(ws, 'A' + str(r), 'Préstamo bancario (€)')
    motor.f(ws, 'B' + str(r), '=B' + str(f_imp), fmt=FMT_EUR)
    f_pr = r
    r += 1
    motor.val(ws, 'A' + str(r), 'Otras fuentes (subvención, ENISA, socio '
              'industrial) (€)')
    motor.val(ws, 'B' + str(r), conf.get('otras_fuentes', 0), fmt=FMT_EUR,
              verde_=True)
    motor.fijar_formato(ws, 'B' + str(r), FMT_EUR)
    f_ot = r
    r += 1
    motor.val(ws, 'A' + str(r), 'TOTAL FINANCIACIÓN (€)', bold=True)
    motor.f(ws, 'B' + str(r),
            '=IF(COUNT(B' + str(f_fp) + ':B' + str(f_ot) + ')=0,"",SUM(B'
            + str(f_fp) + ':B' + str(f_ot) + '))', fmt=FMT_EUR, bold=True)
    f_tot = r
    r += 1
    motor.val(ws, 'A' + str(r), 'Diferencia (necesidad − financiación) (€)',
              bold=True)
    # ROUND obligatorio: sin él la resta de dos sumas de decimales deja
    # −2,3E−10 y el semáforo pinta en rojo un plan que cuadra al céntimo.
    motor.f(ws, 'B' + str(r),
            '=IF(OR(B' + str(f_nec) + '="",B' + str(f_tot) + '=""),"",ROUND(B'
            + str(f_nec) + '-B' + str(f_tot) + ',2))', fmt=FMT_EUR, bold=True)
    motor.permitir_negativo(ws, 'B' + str(r))
    motor.regla_expresion(ws, 'B' + str(r),
                          '=AND(ISNUMBER(B' + str(r) + '),ABS(B' + str(r)
                          + ')>0.01)')
    nota(ws, 'A' + str(r + 1),
         'Si esta fila no es cero, el plan NO está financiado: sube el '
         'préstamo o los fondos propios, baja los meses de colchón o recorta '
         'el CAPEX. Es la primera comprobación que hace un analista de riesgos '
         'y la que decide si el expediente sigue adelante.')
    apunta(cambios, fname, ws, 'bloque «Plan de financiación» en las filas '
           + str(base) + '-' + str(r) + ': necesidad, fondos propios, préstamo, '
           'otras fuentes y cuadre con semáforo (RD-20)')
    return base


def plan_representante(wb, fname, cambios, contenido, subtitulo):
    pl = _pl_mensual_representante(wb, fname, cambios, contenido)
    ws_inv = wb['Inversión']
    fila_cab = motor.fila_cabecera_tabla(ws_inv) or 4
    # ⚠️ Los conceptos de «Inversión» viven en la columna del rótulo, NO en la
    # A (que sólo lleva el «#»). Pasarlo mal dejaba mudos el fondo de maniobra
    # y la tabla de correspondencia, sin un solo aviso (RT-03/RT-04).
    col_etq = _columna_concepto(ws_inv, fila_cab)
    _precargar_inversion(ws_inv, fname, cambios, contenido, 'C', col_etq,
                         fila_cab)
    fin = hoja_financiacion(wb, fname, cambios, contenido, subtitulo)
    hoja_proyeccion(wb, fname, cambios, contenido, subtitulo, pl, fin)
    inv = fondo_de_maniobra(wb, fname, cambios, contenido, pl, 'Inversión',
                            'C', col_etq=col_etq, fila_cab=fila_cab)
    correspondencia_capex(wb, fname, cambios, contenido, 'Inversión', fila_cab,
                          col_etq=col_etq)
    _total_inversion_guardado(ws_inv, fname, cambios, fila_cab, col_etq)
    _desviacion_sin_dato(ws_inv, fname, cambios, 'C', 'D', 'E')
    plan_de_financiacion(wb, fname, cambios, contenido, inv)
    instruccion(wb, 'La pestaña «' + HOJA_PROY + '» ya existe: el Año 1 sale '
                    'del P&L Mensual por referencia, corregido por la rampa '
                    'de arranque de sus dos celdas verdes (los mismos meses '
                    'que modela cash-flow-break-even.xlsx, así que las dos '
                    'facturaciones del año 1 coinciden); los años 2 y 3 salen '
                    'de los dos porcentajes verdes sobre el ritmo de crucero.',
                RX_INSTR_PROY)
    instruccion(wb, 'La columna «Real (€)» de esta hoja «Inversión» se '
                    'entrega VACÍA a propósito: es donde vas anotando lo que '
                    'pagas de verdad, y por eso su fila TOTAL aparece en '
                    'blanco hasta que escribas la primera cifra (el libro '
                    'nunca escribe un 0 donde no hay dato). La «Desviación '
                    '(%)» se enciende sola en cuanto haya un Real.',
                RX_INSTR_REAL)
    instruccion(wb, 'La pestaña «' + HOJA_FIN + '» calcula la cuota del '
                    'préstamo y su cuadro de amortización; sus intereses '
                    'alimentan la proyección y su cuota mensual se repite en '
                    'cash-flow-break-even.xlsx.', RX_INSTR_FIN)
    instruccion(wb, 'Fuente de cada cifra: esta hoja «Inversión» es TU CAPEX '
                    '(el que va al banco). calculadora-capex.xlsx es la hoja '
                    'de RANGOS DE MERCADO, con el desglose por categorías; la '
                    'última columna de «Inversión» dice a qué categoría suya '
                    'corresponde cada concepto.', RX_INSTR_FUENTE)
    return 'inversion-conceptos'


def _total_inversion_guardado(ws, fname, cambios, fila_cab, col_etq):
    """RC-09 — «TOTAL inversión 0,00 €» con la tabla vacía. Misma guarda `COUNT`
    que el resto del paquete (nunca `COUNTIF(rango,"<>")`, ver RT-01)."""
    f_tot = _fila(ws, r'^total', col=col_etq, desde=fila_cab + 1,
                  obligatoria=False)
    if not f_tot:
        return 0
    tocadas = 0
    for c in range(2, ws.max_column + 1):
        from openpyxl.utils import get_column_letter as gcl
        col = gcl(c)
        cel = ws[col + str(f_tot)]
        if cel.data_type != 'f' or 'SUM' not in str(cel.value).upper():
            continue
        rango = col + str(fila_cab + 1) + ':' + col + str(f_tot - 1)
        motor.f(ws, col + str(f_tot),
                '=IF(COUNT(' + rango + ')=0,"",SUM(' + rango + '))',
                fmt=FMT_EUR, bold=True)
        tocadas += 1
    if tocadas:
        apunta(cambios, fname, ws, str(tocadas) + ' totales de la tabla con la '
               'guarda «sin dato = ""» en vez de «0,00 €» (§7-bis.13, RC-09)')
    return tocadas


def _desviacion_sin_dato(ws, fname, cambios, col_prev, col_real, col_desv):
    """§7-bis.13 en la columna de desviación de `Inversión`.

    Representante: `=IF(C5=0,0,(D5-C5)/C5)` pinta **0,0 %** en las 22 filas de
    un libro en blanco, como si el presupuesto cuadrase. Hermanos: `=D5-C5`
    pinta **−110.000 €** de «Diferencia» antes de que el cliente haya escrito
    un solo importe real. Las dos formas mienten con el libro vacío.
    """
    fila_cab = 4
    tocadas = 0
    for r in range(fila_cab + 1, ws.max_row + 1):
        cel = ws[col_desv + str(r)]
        if cel.data_type != 'f':
            continue
        motor.f(ws, col_desv + str(r),
                '=IF(' + col_real + str(r) + '="","",IFERROR((' + col_real
                + str(r) + '-' + col_prev + str(r) + ')/' + col_prev + str(r)
                + ',""))', fmt=FMT_PCT)
        tocadas += 1
    if tocadas:
        # RT-20 · §1.6 nombra «la desviación de CAPEX» entre los sitios donde
        # va el semáforo, y era la ÚNICA columna de desviación del pack sin
        # una sola regla de formato condicional.
        primera = fila_cab + 1
        ultima = max(r for r in range(primera, ws.max_row + 1)
                     if ws[col_desv + str(r)].data_type == 'f')
        rango = col_desv + str(primera) + ':' + col_desv + str(ultima)
        motor.semaforo_isnumber(ws, rango, '$' + col_desv + str(primera),
                                operador='>', umbral='0.1')
        apunta(cambios, fname, ws, str(tocadas) + ' celdas de desviación que '
               'ya no dicen «0,0 %» ni una diferencia falsa con el libro en '
               'blanco (§7-bis.13) + semáforo ISNUMBER sobre ' + rango
               + ' al 10 % (§1.6, RT-20)')
    return tocadas


def plan_hermanos(wb, fname, cambios, contenido, subtitulo):
    pl = _pl_mensual_hermanos(wb, fname, cambios, contenido)
    fin = hoja_financiacion(wb, fname, cambios, contenido, subtitulo)
    hoja_proyeccion(wb, fname, cambios, contenido, subtitulo, pl, fin)
    ws_inv = wb['Inversión']
    f_tot = _fila(ws_inv, r'^total inversion', col=2, obligatoria=False)
    fila_cab_h = motor.fila_cabecera_tabla(ws_inv) or 4
    fondo_de_maniobra(wb, fname, cambios, contenido, pl, 'Inversión', 'C',
                      col_etq=_columna_concepto(ws_inv, fila_cab_h),
                      fila_cab=fila_cab_h)
    _diferencia_hermanos(ws_inv, fname, cambios)
    instruccion(wb, 'La pestaña «' + HOJA_PROY + '» ya existe: el Año 1 sale '
                    'del P&L Mensual por referencia, corregido por la rampa '
                    'de arranque de sus dos celdas verdes; los años 2 y 3, '
                    'del ritmo de crucero.', RX_INSTR_PROY)
    instruccion(wb, 'La pestaña «' + HOJA_FIN + '» calcula la cuota del '
                    'préstamo y su cuadro de amortización.', RX_INSTR_FIN)
    return 'inversion-categorias', f_tot


def _diferencia_hermanos(ws, fname, cambios):
    """`E5='=D5-C5'` → `''` mientras `D` esté vacía (§7-bis.13)."""
    tocadas = 0
    for r in range(5, ws.max_row + 1):
        cel = ws['E' + str(r)]
        if cel.data_type != 'f' or 'SUM' in str(cel.value).upper():
            continue
        motor.f(ws, 'E' + str(r),
                '=IF(D' + str(r) + '="","",D' + str(r) + '-C' + str(r) + ')',
                fmt=FMT_EUR)
        tocadas += 1
    if tocadas:
        apunta(cambios, fname, ws, str(tocadas) + ' celdas de «Diferencia» que '
               'dejaban de mentir un −110.000 € con la columna Real vacía')
    return tocadas


def plan_panaderia(wb, fname, cambios, contenido, subtitulo):
    """NUEVO-02 (§2.5) — el EBITDA que vale exactamente la facturación.

    `B23` sumaba dos veces `B14:B16` y se dejaba fuera `B21`; `B24` restaba una
    fila **vacía** (`B22`), así que el «EBITDA» cacheado era 300.000 €, la
    FACTURACIÓN TOTAL; y `B25` dividía el total de COSTES entre la facturación
    y publicaba un margen del 122,97 % junto a una nota que dice «15-22 %
    objetivo». Las tres fórmulas se rehacen por estructura: del primer bloque de
    coste a la última línea de coste, sin saltarse ninguna ni repetir ninguna.
    """
    ws = wb['P&L 3 años']
    f_fact = _fila(ws, r'^facturacion total', fname=fname)
    f_var = _fila(ws, RX_BLOQUE_VAR, fname=fname)
    f_tot = _fila(ws, r'^total costes', fname=fname)
    f_eb = _fila(ws, RX_EBITDA, desde=f_tot, fname=fname)
    f_pct = _fila(ws, RX_PCT_EBITDA, desde=f_eb, fname=fname)
    ultima = _ultimo_detalle(ws, f_var + 1, f_tot - 1, (RX_BLOQUE_FIJ,))
    cols = ['B', 'C', 'D']
    for col in cols:
        a_formula(ws, col + str(f_tot),
                  '=SUM(' + col + str(f_var + 1) + ':' + col + str(ultima)
                  + ')', fname, cambios, fmt=FMT_EUR,
                  nota_dif='NUEVO-02: sumaba dos veces el bloque de energía y '
                           'se dejaba fuera «Seguros + tasas + amortización»')
        a_formula(ws, col + str(f_eb),
                  '=' + col + str(f_fact) + '-' + col + str(f_tot), fname,
                  cambios, fmt=FMT_EUR,
                  nota_dif='NUEVO-02: restaba una fila VACÍA, así que el '
                           'EBITDA era la facturación entera')
        a_formula(ws, col + str(f_pct),
                  '=IF(' + col + str(f_fact) + '=0,"",' + col + str(f_eb) + '/'
                  + col + str(f_fact) + ')', fname, cambios, fmt=FMT_PCT,
                  nota_dif='NUEVO-02: dividía el TOTAL DE COSTES entre la '
                           'facturación y publicaba 122,97 %')
    # §1.4, falso positivo MEDIDO: la etiqueta «Personal (5-7 personas)» casa
    # con el patrón de RECUENTO del motor (`personas`) y su columna se
    # reformatea a `#,##0`, así que 118.000 € de nómina se imprimen sin el
    # euro. Las filas de dinero de esta hoja se CLAVAN en euros.
    clavadas = 0
    for r in range(f_fact, f_eb + 1):
        if r in (f_pct,):
            continue
        for col in cols:
            cel = ws[col + str(r)]
            if cel.value is None or isinstance(cel.value, str) \
                    and cel.data_type != 'f':
                continue
            motor.fijar_formato(ws, col + str(r), FMT_EUR)
            clavadas += 1
    for col in cols:
        motor.fijar_formato(ws, col + str(f_pct), FMT_PCT)
    apunta(cambios, fname, ws, str(clavadas) + ' celdas de importe clavadas en '
           'euros frente a la regla de columna del §1.4 (la etiqueta «Personal '
           '(5-7 personas)» las llevaba a #,##0)')
    _parametrizar_porcentajes(ws, cols, fname, cambios)
    _semaforo_negativo(ws, cols, (f_eb, f_pct))
    apunta(cambios, fname, ws, 'NUEVO-02 corregido: TOTAL COSTES = SUM('
           + str(f_var + 1) + ':' + str(ultima) + '), EBITDA = facturación − '
           'costes, margen = EBITDA / facturación')
    return 'pl-3-anos'


RX_PCT_INCRUSTADO = re.compile(r'^=\s*([A-Z]{1,2}\d{1,4})\s*\*\s*'
                               r'(0?\.\d+)\s*$')


def _etiqueta_pct(texto):
    """B-04 — la etiqueta del bloque de parámetros tiene que DECIR que es un
    porcentaje, o el motor la reclasifica a euros en la 2.ª pasada.

    Medido en panadería: `A28 = 'Coste materias primas (harinas +
    ingredientes)'` contiene «Coste» y no contiene «(%)», así que
    `motor.formato_por_etiqueta` la leía como importe y `B28/C28/D28` pasaban
    de `0.0%` a `#,##0.00 €` — el 0,20 dejaba de leerse «20,0 %» y pasaba a
    «0,20 €» en el libro que se lleva al banco. La 1.ª pasada aguantaba sólo
    porque `_parametrizar_porcentajes` clava el formato con `fijar_formato`; en
    la 2.ª ya no hay literal que extraer, la función sale por el `return 0` de
    arriba y el pin no se vuelve a poner. Con el «(%)» al final la etiqueta se
    defiende sola —en las dos pasadas y sin depender de ningún pin— y la DV
    tampoco se parte en dos grupos.
    """
    t = str(texto).strip()
    return t if motor.RX_PCT.search(t) else (t + ' (%)')


def _parametrizar_porcentajes(ws, cols, fname, cambios):
    """Convención de familia: **parámetro en celda, nunca literal dentro de la
    fórmula** (§1 «Convenciones», §1.5).

    Caso medido en panadería: `'P&L 3 años'!B12='=B9*0.20'`, `C12='=C9*0.19'`,
    `D12='=D9*0.18'` (materias primas), y lo mismo con el 0,015 del packaging y
    el 0,025 de la energía: **9 tasas** que el cliente no puede tocar sin
    reescribir la fórmula, en el libro que se lleva al banco. Se bajan a un
    bloque de celdas verdes con la misma rejilla de años y la fórmula pasa a
    referenciarlas. El valor no cambia: se conserva el que había.
    """
    filas = {}
    for r in range(1, ws.max_row + 1):
        for col in cols:
            cel = ws[col + str(r)]
            if cel.data_type != 'f':
                continue
            m = RX_PCT_INCRUSTADO.match(str(cel.value).replace(' ', ''))
            if m:
                filas.setdefault(r, {})[col] = (m.group(1), float(m.group(2)))
    if not filas:
        return 0
    base = base_bloque(ws, r'^parametros en % sobre la facturacion')
    motor.val(ws, 'A' + str(base),
              'PARÁMETROS EN % SOBRE LA FACTURACIÓN', bold=True)
    nota(ws, 'E' + str(base),
         'Estaban escritos DENTRO de las fórmulas de arriba: para cambiar un '
         'food cost había que editar la fórmula. Ahora se cambian aquí.')
    tocadas = []
    for i, r in enumerate(sorted(filas)):
        destino = base + 1 + i
        motor.val(ws, 'A' + str(destino),
                  _etiqueta_pct(ws['A' + str(r)].value
                                or ('Fila ' + str(r))))
        for col, (ref, pct) in filas[r].items():
            motor.val(ws, col + str(destino), pct, fmt=FMT_PCT, verde_=True)
            motor.fijar_formato(ws, col + str(destino), FMT_PCT)
            motor.f(ws, col + str(r),
                    '=' + ref + '*' + col + str(destino), fmt=FMT_EUR)
            tocadas.append(col + str(r) + '→' + col + str(destino) + '='
                           + str(pct))
    apunta(cambios, fname, ws, str(len(tocadas)) + ' porcentajes sacados de '
           'dentro de las fórmulas a celdas verdes en las filas '
           + str(base + 1) + '-' + str(base + len(filas)) + ': '
           + str(tocadas))
    return len(tocadas)


def plan_financiero(wb, fname, cambios, contenido, subtitulo):
    variante = variante_plan(wb, fname)
    if variante == 'inversion-conceptos':
        plan_representante(wb, fname, cambios, contenido, subtitulo)
    elif variante == 'inversion-categorias':
        plan_hermanos(wb, fname, cambios, contenido, subtitulo)
    else:
        plan_panaderia(wb, fname, cambios, contenido, subtitulo)
    return variante


# ==========================================================================
# §2.4 · cash-flow-break-even.xlsx  (TEC-03, DOM-09, COM-06, DOM-03, DOM-22)
# ==========================================================================
ET_NETO_IVA = 'FLUJO DE CAJA NETO CON IVA Y DEUDA'
ET_ACUM_IVA = 'FLUJO ACUMULADO CON IVA Y DEUDA'
ET_BE_MES = 'Mes de break-even de caja'


def variante_cash(wb, fname=''):
    hojas = set(wb.sheetnames)
    if 'Cash Flow 12 Meses' in hojas:
        return '12-meses-total', 'Cash Flow 12 Meses', None
    if 'Cash Flow' in hojas and 'Break-Even' in hojas:
        return 'cash-y-breakeven', 'Cash Flow', 'Break-Even'
    caja = [h for h in wb.sheetnames if _norm(h).startswith('cash flow')]
    if caja:
        return 'cash-plurianual', caja[0], None
    raise VarianteDesconocida(
        fname + ': libro de cash flow no reconocido. Hojas=' + repr(
            wb.sheetnames))


def _bloque_iva_y_deuda(ws, meses, total, fila_neto, fila_acum, cambios, fname,
                        contenido):
    """Las tres filas de IVA del modelo 303 + la cuota del préstamo (§2.4), la
    tesorería de la apertura (RD-07) y el IVA de la bebida alcohólica (RD-14).

    El cash flow va **con IVA porque es caja**, y el IVA soportado del CAPEX es
    tesorería adelantada que no aparecía por ningún lado. Se monta como una
    CAPA sobre la fila del flujo neto, sin insertar filas: `insert_rows` no
    reescribe las fórmulas y partiría los `SUM` horizontales que ya existen.
    """
    base = base_bloque(ws, r'^iva \(modelo 303\) y servicio de la deuda$')
    conf = ((getattr(contenido, 'CASH', None) or {}) if contenido else {})
    motor.val(ws, 'A' + str(base),
              'IVA (MODELO 303) Y SERVICIO DE LA DEUDA', bold=True)
    celda_iva = motor.escribir_parametro(ws, base + 1, 'A', 'B',
                                         'iva_restauracion')
    motor.fijar_formato(ws, celda_iva, FMT_PCT)
    # ABSOLUTA: la misma celda de tipo alimenta las doce columnas de mes, y una
    # referencia relativa se rompería en cuanto el cliente arrastrase la fila.
    celda_iva = '$' + celda_iva[0] + '$' + celda_iva[1:]
    # RD-14 · el 10 % único sobre TODA la facturación deja corta la liquidación
    # del 303: la bodega de un gastronómico con maridaje es una parte grande de
    # la venta y va al 21 %. El propio libro lo decía en la nota de al lado y
    # la fórmula no lo hacía.
    motor.val(ws, 'A' + str(base + 2),
              'Tipo de IVA de la bebida alcohólica (%)')
    motor.val(ws, 'B' + str(base + 2), conf.get('iva_bebida', 0.21),
              fmt=FMT_PCT, verde_=True)
    motor.fijar_formato(ws, 'B' + str(base + 2), FMT_PCT)
    celda_iva_beb = '$B$' + str(base + 2)
    # §1.3 — ningún literal dentro de una fórmula: el tipo de los gastos y la
    # base de gasto con IVA son parámetros del cliente, no constantes.
    iva_conf0 = ((conf.get('mes_tipo') or {}).get('iva_soportado') or {})
    motor.val(ws, 'A' + str(base + 3),
              'Tipo de IVA general de los gastos (%)')
    motor.val(ws, 'B' + str(base + 3), iva_conf0.get('tipo_fijos', 0.21),
              fmt=FMT_PCT, verde_=True)
    motor.fijar_formato(ws, 'B' + str(base + 3), FMT_PCT)
    celda_iva_gas = '$B$' + str(base + 3)
    motor.val(ws, 'A' + str(base + 4), 'Base mensual de gastos con IVA '
              '(alquiler, suministros, marketing, otros) (€)')
    motor.val(ws, 'B' + str(base + 4), iva_conf0.get('fijos_con_iva'),
              fmt=FMT_EUR, verde_=True)
    motor.fijar_formato(ws, 'B' + str(base + 4), FMT_EUR)
    celda_base_gas = '$B$' + str(base + 4)
    nota(ws, 'A' + str(base + 5),
         '10 % general de restauración y 21 % en bebidas alcohólicas (art. 91 '
         'de la Ley del IVA). Las nóminas y la Seguridad Social NO llevan IVA: '
         'por eso la base de gastos con IVA no es el total de gastos. Las '
         'filas de arriba se escriben SIN IVA, igual que el P&L; esta capa lo '
         'añade porque el cash flow es caja.')
    filas = {
        'repercutido': (base + 6, '(+) IVA repercutido (cobrado con las '
                                  'ventas)'),
        'soportado': (base + 7, '(-) IVA soportado (compras y gastos con IVA; '
                                'las nóminas y la Seguridad Social no llevan)'),
        'liquidacion': (base + 8, '(-) Liquidación de IVA (modelo 303) — '
                                  'trimestre natural, se ingresa del 1 al 20 '
                                  'del mes siguiente'),
        'cuota': (base + 9, '(-) Cuota del préstamo (capital + intereses) — la '
                            'calcula plan-financiero-3-anos.xlsx, hoja «'
                            + HOJA_FIN + '»'),
        'neto': (base + 10, ET_NETO_IVA),
        'acum': (base + 11, ET_ACUM_IVA),
        'q4': (base + 12, 'IVA del 4.º trimestre — se liquida en enero del año '
                          'siguiente (no es caja de este ejercicio)'),
    }
    for clave, (r, etiqueta) in filas.items():
        motor.val(ws, 'A' + str(r), etiqueta,
                  bold=clave in ('neto', 'acum'))
    # RC-07/RD-21 · la nota afirmaba que el break-even se lee CON la cuota del
    # préstamo y la fila de la cuota estaba vacía en los 12 meses. Y estos 12
    # meses son el AÑO 1, que con el préstamo de ejemplo es el año de CARENCIA:
    # se pagan sólo intereses, no la cuota completa.
    cuota = conf.get('cuota_carencia') if conf.get('anio') == 1 \
        else conf.get('cuota_mensual')
    f_rep, f_sop = filas['repercutido'][0], filas['soportado'][0]
    f_liq, f_cuo = filas['liquidacion'][0], filas['cuota'][0]
    f_net, f_acu = filas['neto'][0], filas['acum'][0]
    f_beb = _fila(ws, r'^bebidas y vinos|^bebidas$|^vinos y bebidas',
                  hasta=fila_neto['ingresos'], obligatoria=False)
    mes_tipo = conf.get('mes_tipo') or {}
    iva_conf = mes_tipo.get('iva_soportado') or {}
    f_mp = _fila(ws, r'^materia prima', obligatoria=False)

    for i, col in enumerate(meses):
        if f_beb:
            rep = ('=(' + col + str(fila_neto['ingresos']) + '-IF(ISNUMBER('
                   + col + str(f_beb) + '),' + col + str(f_beb) + ',0))*'
                   + celda_iva + '+IF(ISNUMBER(' + col + str(f_beb) + '),'
                   + col + str(f_beb) + ',0)*' + celda_iva_beb)
        else:
            rep = '=' + col + str(fila_neto['ingresos']) + '*' + celda_iva
        motor.f(ws, col + str(f_rep), rep, fmt=FMT_EUR)
        # IVA soportado: el 10 % de la materia prima y el 21 % de los gastos
        # con IVA. Si el módulo de contenido no dice cuáles son, queda en
        # verde para que lo escriba el cliente (no se inventa una base).
        if f_mp and iva_conf.get('fijos_con_iva') is not None:
            motor.f(ws, col + str(f_sop),
                    '=IF(ISNUMBER(' + col + str(f_mp) + '),' + col + str(f_mp)
                    + '*' + celda_iva + ',0)+IF(ISNUMBER(' + celda_base_gas
                    + '),' + celda_base_gas + '*' + celda_iva_gas + ',0)',
                    fmt=FMT_EUR)
        else:
            motor.val(ws, col + str(f_sop), None, fmt=FMT_EUR)
            motor.verde(ws, col + str(f_sop))
        motor.val(ws, col + str(f_cuo), cuota, fmt=FMT_EUR)
        motor.verde(ws, col + str(f_cuo))
        # liquidación trimestral: abril, julio y octubre liquidan el trimestre
        # natural anterior (el 4.º cae fuera de los 12 meses y va en su fila)
        if i in (3, 6, 9):
            ini, fin = meses[i - 3], meses[i - 1]
            motor.f(ws, col + str(f_liq),
                    '=SUM(' + ini + str(f_rep) + ':' + fin + str(f_rep)
                    + ')-SUM(' + ini + str(f_sop) + ':' + fin + str(f_sop)
                    + ')', fmt=FMT_EUR)
        motor.f(ws, col + str(f_net),
                '=' + col + str(fila_neto['neto']) + '+' + col + str(f_rep)
                + '-' + col + str(f_sop) + '-' + ('IF(ISNUMBER(' + col
                + str(f_liq) + '),' + col + str(f_liq) + ',0)') + '-'
                + col + str(f_cuo), fmt=FMT_EUR, bold=True)
        if i == 0:
            motor.f(ws, col + str(f_acu), '=' + col + str(f_net), fmt=FMT_EUR,
                    bold=True)
        else:
            ant = meses[i - 1]
            motor.f(ws, col + str(f_acu),
                    '=' + ant + str(f_acu) + '+' + col + str(f_net),
                    fmt=FMT_EUR, bold=True)
    ultimo, primero = meses[-1], meses[0]
    motor.f(ws, (total or ultimo) + str(filas['q4'][0]),
            '=SUM(' + meses[9] + str(f_rep) + ':' + ultimo + str(f_rep)
            + ')-SUM(' + meses[9] + str(f_sop) + ':' + ultimo + str(f_sop)
            + ')', fmt=FMT_EUR)
    if total:
        for r in (f_rep, f_sop, f_liq, f_cuo, f_net):
            motor.f(ws, total + str(r),
                    '=SUM(' + primero + str(r) + ':' + ultimo + str(r) + ')',
                    fmt=FMT_EUR)
    motor.permitir_negativo(ws, primero + str(f_acu) + ':' + ultimo
                            + str(f_acu))
    motor.semaforo_isnumber(ws, primero + str(f_acu) + ':' + ultimo
                            + str(f_acu), primero + str(f_acu))
    # RT-21/RC-30 · la cuota NO se escribe como cifra dentro de una nota: en
    # cuanto el cliente cambie importe, plazo, tipo o carencia —que es lo que
    # la hoja le pide hacer— la nota seguiría diciendo la vieja.
    nota(ws, 'A' + str(filas['q4'][0] + 1),
         'La cuota la calcula plan-financiero-3-anos.xlsx, hoja «' + HOJA_FIN
         + '»: durante la carencia se pagan SÓLO intereses (su celda «Cuota '
         'mensual DURANTE la carencia») y después la cuota completa (su celda '
         '«Cuota MENSUAL tras la carencia»). Estos 12 meses son el AÑO 1: con '
         'el préstamo de ejemplo caen dentro de la carencia. Copia ahí el '
         'valor que te corresponda.')

    # ---- RD-07 · tesorería de la apertura --------------------------------
    ap = conf.get('apertura') or {}
    r = filas['q4'][0] + 2
    motor.val(ws, 'A' + str(r), 'TESORERÍA DE LA APERTURA', bold=True)
    f_saldo = r + 1
    motor.val(ws, 'A' + str(f_saldo), 'Saldo inicial de tesorería (€) — sólo '
              'el mes 1')
    motor.val(ws, primero + str(f_saldo), ap.get('saldo_inicial'), fmt=FMT_EUR,
              verde_=True)
    motor.fijar_formato(ws, primero + str(f_saldo), FMT_EUR)
    f_capex = r + 2
    motor.val(ws, 'A' + str(f_capex), '(-) Desembolso de la inversión (CAPEX) '
              'pagado dentro de estos 12 meses (€)')
    f_ivacap = r + 3
    motor.val(ws, 'A' + str(f_ivacap), '(-) IVA soportado de la inversión '
              '(recuperable en el 303, con su plazo) (€)')
    f_tes = r + 4
    motor.val(ws, 'A' + str(f_tes), 'SALDO DE TESORERÍA (fin de mes)',
              bold=True)
    for i, col in enumerate(meses):
        for f_ in (f_capex, f_ivacap):
            motor.val(ws, col + str(f_), ap.get('desembolso_capex')
                      if f_ == f_capex else ap.get('iva_capex'), fmt=FMT_EUR)
            motor.verde(ws, col + str(f_))
            motor.fijar_formato(ws, col + str(f_), FMT_EUR)
        arranque = (primero + str(f_saldo)) if i == 0 \
            else (meses[i - 1] + str(f_tes))
        motor.f(ws, col + str(f_tes),
                '=IF(ISNUMBER(' + arranque + '),' + arranque + ',0)+' + col
                + str(f_net) + '-IF(ISNUMBER(' + col + str(f_capex) + '),'
                + col + str(f_capex) + ',0)-IF(ISNUMBER(' + col + str(f_ivacap)
                + '),' + col + str(f_ivacap) + ',0)', fmt=FMT_EUR, bold=True)
    motor.permitir_negativo(ws, primero + str(f_tes) + ':' + ultimo
                            + str(f_tes))
    motor.semaforo_isnumber(ws, primero + str(f_tes) + ':' + ultimo
                            + str(f_tes), primero + str(f_tes))
    nota(ws, 'A' + str(f_tes + 1),
         'El «flujo acumulado» de arriba arranca en cero: mide la EXPLOTACIÓN. '
         'Esta fila mide la CAJA, y por eso parte del saldo con el que abres '
         '—el fondo de maniobra que dimensiona plan-financiero-3-anos.xlsx, '
         'hoja «Inversión»— y resta la parte de la inversión y de su IVA que '
         'pagues dentro de estos doce meses. Un cash flow de apertura sin '
         'saldo inicial dice que el negocio va bien el primer mes en que gana '
         'dinero, no cuando ha recuperado lo invertido.')
    apunta(cambios, fname, ws, 'capa de IVA (303, con el 21 % de la bebida '
           'alcohólica), cuota del préstamo y tesorería de la apertura en las '
           'filas ' + str(f_rep) + '-' + str(f_tes) + ' (DOM-03, DOM-22, §2.4, '
           'RD-07, RD-14, RD-21)')
    return {'neto': f_net, 'acum': f_acu, 'liquidacion': f_liq,
            'cuota': f_cuo, 'iva': celda_iva, 'tesoreria': f_tes,
            'saldo': f_saldo}


def _bloque_break_even(ws, capa, meses, cambios, fname, contenido):
    """Break-even en MESES y en € (TEC-03, DOM-09, COM-06) — ahora con la cuota.

    RD-05: el umbral y los cubiertos/día se calculaban SÓLO con los costes
    fijos de explotación, justo debajo de una nota que promete lo contrario
    («un punto de equilibrio sin la cuota del préstamo es un punto de
    equilibrio falso»). Se publican las DOS líneas, que es lo que un consultor
    lleva al banco: el umbral de explotación y el umbral con servicio de deuda.
    """
    conf = ((getattr(contenido, 'CASH', None) or {}) if contenido else {})
    be = conf.get('break_even') or {}
    cuota = (conf.get('cuota_carencia') if conf.get('anio') == 1
             else conf.get('cuota_mensual'))
    base = base_bloque(ws, r'^break-even$')
    motor.val(ws, 'A' + str(base), 'BREAK-EVEN', bold=True)
    f_fij, f_mc = base + 1, base + 2
    f_umb, f_cuo, f_umbd = base + 3, base + 4, base + 5
    f_tk, f_dias, f_cub, f_mes = base + 6, base + 7, base + 8, base + 9
    entradas = ((f_fij, 'Costes fijos mensuales, sin amortización (€)',
                 be.get('costes_fijos'), FMT_EUR),
                (f_mc, 'Margen de contribución (%)',
                 be.get('margen_contribucion'), FMT_PCT),
                (f_cuo, 'Cuota mensual del préstamo (€)', cuota, FMT_EUR),
                (f_tk, 'Ticket medio (€)', be.get('ticket_medio'), FMT_EUR),
                (f_dias, 'Días abierto/mes', be.get('dias_mes'), FMT_ENT))
    for fila, etiqueta, valor, fmt in entradas:
        motor.val(ws, 'A' + str(fila), etiqueta)
        motor.val(ws, 'B' + str(fila), valor, fmt=fmt, verde_=True)
        motor.fijar_formato(ws, 'B' + str(fila), fmt)
    motor.val(ws, 'A' + str(f_umb),
              'Umbral de ventas mensual — sólo explotación (€)')
    motor.f(ws, 'B' + str(f_umb),
            '=IFERROR(B' + str(f_fij) + '/B' + str(f_mc) + ',"")', fmt=FMT_EUR,
            bold=True)
    motor.fijar_formato(ws, 'B' + str(f_umb), FMT_EUR)
    motor.val(ws, 'A' + str(f_umbd),
              'Umbral de ventas mensual CON el servicio de la deuda (€)')
    motor.f(ws, 'B' + str(f_umbd),
            '=IFERROR((B' + str(f_fij) + '+IF(ISNUMBER(B' + str(f_cuo) + '),B'
            + str(f_cuo) + ',0))/B' + str(f_mc) + ',"")', fmt=FMT_EUR,
            bold=True)
    motor.fijar_formato(ws, 'B' + str(f_umbd), FMT_EUR)
    motor.val(ws, 'A' + str(f_cub),
              'Cubiertos/día necesarios (con el servicio de la deuda)')
    motor.f(ws, 'B' + str(f_cub),
            '=IFERROR(B' + str(f_umbd) + '/(B' + str(f_tk) + '*B'
            + str(f_dias) + '),"")', fmt=FMT_ENT, bold=True)
    motor.fijar_formato(ws, 'B' + str(f_cub), FMT_ENT)
    motor.val(ws, 'A' + str(f_mes), ET_BE_MES)
    motor.f(ws, 'B' + str(f_mes),
            '=IFERROR(MATCH(TRUE,INDEX(' + meses[0] + str(capa['acum']) + ':'
            + meses[-1] + str(capa['acum']) + '>0,0),0),"No alcanzado")',
            bold=True)
    motor.fijar_formato(ws, 'B' + str(f_mes), 'General')
    nota(ws, 'A' + str(f_mes + 1),
         'El mes de break-even se lee sobre el FLUJO ACUMULADO CON IVA Y '
         'DEUDA: un punto de equilibrio sin la cuota del préstamo es un punto '
         'de equilibrio falso. Si el acumulado nunca cruza a positivo, la '
         'celda dice «No alcanzado» — no un número. Los dos umbrales de arriba '
         'son la misma idea en euros: el de explotación paga la estructura; el '
         'de la deuda paga además al banco, y es el que hay que superar para '
         'no comerse el fondo de maniobra.')
    apunta(cambios, fname, ws, 'break-even en meses y en € en las filas '
           + str(f_fij) + '-' + str(f_mes) + ', con umbral de explotación Y '
           'umbral con servicio de deuda (TEC-03, DOM-09, COM-06, RD-05)')
    return base


def cash_plurianual(wb, fname, cambios, contenido, nombre_caja):
    """Panadería: `Cash Flow 24m`, un molde propio que YA calcula los flujos.

    Rejilla medida: `Ingresos` · `Coste materias primas` · `Personal` ·
    `Alquiler…` (las salidas van en NEGATIVO, no en un bloque de gastos) ·
    `Flujo operativo` = `SUM` · `CAPEX inicial` · `Flujo neto` · `Caja
    acumulada`, con columnas `M1…M12` + `Año 1` + `Año 2`. No tiene filas de
    «Total ingresos» ni «Total gastos», así que el camino del §2.4 no le vale:
    lo que le falta es **el break-even**, que hoy es el TEXTO
    «M10-M12 (realista)» tecleado en una celda — una estimación disfrazada de
    resultado, en el fichero que se llama `cash-flow-break-even`.
    """
    ws = wb[nombre_caja]
    fila_cab = motor.fila_cabecera_tabla(ws) or 3
    meses, _total = columnas_mes(ws, fila_cab)
    meses = meses[:12]
    f_acum = _fila(ws, r'^caja acumulada|^saldo acumulado|^flujo acumulado',
                   fname=fname)
    f_be = _fila(ws, r'^break-?even', obligatoria=False)
    if f_be is None:
        f_be = base_bloque(ws, r'^mes de break-?even')
        motor.val(ws, 'A' + str(f_be), 'Mes de break-even de caja')
    motor.f(ws, 'B' + str(f_be),
            '=IFERROR(MATCH(TRUE,INDEX(' + meses[0] + str(f_acum) + ':'
            + meses[-1] + str(f_acum) + '>0,0),0),"No alcanzado")', bold=True)
    motor.limpiar_rango(ws, 'C' + str(f_be))
    nota(ws, 'C' + str(f_be),
         'Calculado sobre la fila «' + str(ws['A' + str(f_acum)].value)
         + '»: el primer mes en que la caja cruza a positivo, buscando SÓLO en '
           'las doce columnas mensuales (las de Año 1 y Año 2 son acumulados '
           'anuales, no meses). Si no cruza en esos doce meses dice «No '
           'alcanzado» — antes había aquí una estimación escrita a mano.')
    apunta(cambios, fname, ws, 'break-even calculado en B' + str(f_be)
           + ' sobre la fila ' + str(f_acum) + ', en lugar del texto fijo '
             '«M10-M12 (realista)» (TEC-03, DOM-09)')
    apunta(cambios, fname, ws, 'PENDIENTE T6/panadería: esta rejilla lleva las '
           'salidas en NEGATIVO y va a 24 meses, así que la capa de IVA (303) '
           'y la cuota del préstamo del §2.4 necesitan su propio mapa en '
           'contenido_guia_panaderia_obrador/a.py; no se aplica a ciegas.')
    return 'cash-plurianual'


def cash_flow(wb, fname, cambios, contenido, subtitulo):
    variante, nombre_caja, nombre_be = variante_cash(wb, fname)
    if variante == 'cash-plurianual':
        return cash_plurianual(wb, fname, cambios, contenido, nombre_caja)
    ws = wb[nombre_caja]
    fila_cab = motor.fila_cabecera_tabla(ws) or 4
    meses, total = columnas_mes(ws, fila_cab)
    if len(meses) < 12:
        raise VarianteDesconocida(
            fname + ':' + ws.title + ': ' + str(len(meses)) + ' columnas de '
            'mes leídas en la cabecera de la fila ' + str(fila_cab)
            + ' (se esperaban 12)')
    meses = meses[:12]

    f_ing_bloque = _fila(ws, r'^ingresos$|^entradas$', fname=fname)
    f_tot_ing = _fila(ws, r'^total ingresos$|^total entradas$', fname=fname)
    f_gas_bloque = _fila(ws, r'^gastos$|^salidas$', fname=fname)
    f_tot_gas = _fila(ws, r'^total gastos$|^total salidas$', fname=fname)
    f_neto = _fila(ws, r'^flujo de caja neto$|^flujo neto mensual$',
                   fname=fname)
    f_acum = _fila(ws, r'^flujo acumulado$|^saldo acumulado$', fname=fname)
    f_saldo = _fila(ws, r'^saldo inicial$', obligatoria=False)

    for i, col in enumerate(meses):
        motor.f(ws, col + str(f_tot_ing),
                '=SUM(' + col + str(f_ing_bloque + 1) + ':' + col
                + str(f_tot_ing - 1) + ')', fmt=FMT_EUR)
        motor.f(ws, col + str(f_tot_gas),
                '=SUM(' + col + str(f_gas_bloque + 1) + ':' + col
                + str(f_tot_gas - 1) + ')', fmt=FMT_EUR)
        motor.f(ws, col + str(f_neto),
                '=' + col + str(f_tot_ing) + '-' + col + str(f_tot_gas),
                fmt=FMT_EUR, bold=True)
        if i == 0:
            arranque = ('=' + col + str(f_saldo) + '+' + col + str(f_neto)) \
                if f_saldo else ('=' + col + str(f_neto))
            motor.f(ws, col + str(f_acum), arranque, fmt=FMT_EUR, bold=True)
        else:
            ant = meses[i - 1]
            motor.f(ws, col + str(f_acum),
                    '=' + ant + str(f_acum) + '+' + col + str(f_neto),
                    fmt=FMT_EUR, bold=True)
            if f_saldo:
                motor.f(ws, col + str(f_saldo), '=' + ant + str(f_acum),
                        fmt=FMT_EUR)
    if total:
        for r in (f_tot_ing, f_tot_gas, f_neto):
            motor.f(ws, total + str(r),
                    '=SUM(' + meses[0] + str(r) + ':' + meses[-1] + str(r)
                    + ')', fmt=FMT_EUR)
    motor.semaforo_isnumber(ws, meses[0] + str(f_acum) + ':' + meses[-1]
                            + str(f_acum), meses[0] + str(f_acum))
    apunta(cambios, fname, ws, 'totales, flujo neto y flujo acumulado en las '
           'filas ' + str(f_tot_ing) + '/' + str(f_tot_gas) + '/' + str(f_neto)
           + '/' + str(f_acum) + ' sobre ' + str(len(meses))
           + ' meses (TEC-03, DOM-09)')

    _precargar_cash(ws, meses, cambios, fname, contenido,
                    (f_ing_bloque, f_tot_ing), (f_gas_bloque, f_tot_gas))
    capa = _bloque_iva_y_deuda(ws, meses, total,
                               {'ingresos': f_tot_ing, 'neto': f_neto},
                               f_acum, cambios, fname, contenido)
    destino = wb[nombre_be] if nombre_be else ws
    if nombre_be:
        _break_even_hoja_propia(destino, capa, meses, ws.title, cambios, fname,
                                contenido)
    else:
        _bloque_break_even(ws, capa, meses, cambios, fname, contenido)
    instruccion(wb, 'El break-even se lee sobre el flujo acumulado CON IVA y '
                    'con la cuota del préstamo: sin la cuota, el punto de '
                    'equilibrio sale antes de tiempo.',
                re.compile(r'^El break-even se lee'))
    return variante


def _precargar_cash(ws, meses, cambios, fname, contenido, bloque_ing,
                    bloque_gas):
    """RD-06/RC-08 — las 12 columnas se entregaban VACÍAS y el punto de
    equilibrio decía «No alcanzado» nada más abrir el fichero, en la plantilla
    que la tarjeta vende como «punto de equilibrio automático».

    Se precarga el escenario REALISTA con una rampa de apertura (§7-bis.7: la
    misma facturación que el P&L), en VERDE y como valor de ejemplo (§1.2). Si
    el módulo de contenido no la trae, la hoja se queda vacía: aquí no se
    inventa ninguna serie.
    """
    conf = ((getattr(contenido, 'CASH', None) or {}) if contenido else {})
    mes_tipo = conf.get('mes_tipo') or {}
    rampa = conf.get('rampa')
    if not mes_tipo or not rampa:
        return 0
    grupos = ((mes_tipo.get('ingresos') or {}, bloque_ing, True),
              (mes_tipo.get('variables') or {}, bloque_gas, True),
              (mes_tipo.get('fijos') or {}, bloque_gas, False))
    puestas = 0
    for valores, (ini, fin), con_rampa in grupos:
        for patron, valor in valores.items():
            r = _fila(ws, patron, desde=ini + 1, hasta=fin - 1,
                      obligatoria=False)
            if r is None:
                continue
            for i, col in enumerate(meses):
                if ws[col + str(r)].data_type == 'f':
                    continue
                factor = rampa[i] if (con_rampa and i < len(rampa)) else 1.0
                motor.val(ws, col + str(r), round(valor * factor, 2),
                          fmt=FMT_EUR, verde_=True)
                motor.fijar_formato(ws, col + str(r), FMT_EUR)
                puestas += 1
    if puestas:
        apunta(cambios, fname, ws, str(puestas) + ' celdas de los 12 meses '
               'precargadas en verde con el escenario REALISTA y una rampa de '
               'apertura (' + ' · '.join('%d %%' % round(x * 100)
                                         for x in rampa)
               + '): el punto de equilibrio deja de decir «No alcanzado» al '
               'abrir el fichero (RD-06, RC-08)')
    return puestas


def _break_even_hoja_propia(ws, capa, meses, hoja_caja, cambios, fname,
                            contenido):
    """Hermanos: la hoja `Break-Even` YA existe con `B11` y `B12` y deja
    **vacía justo `B13` «Break-Even (meses)»** — la celda que da nombre al
    fichero."""
    conf = ((getattr(contenido, 'CASH', None) or {}).get('break_even')
            if contenido else None) or {}
    f_fact = _fila(ws, r'^facturacion mensual estimada', obligatoria=False)
    f_margen = _fila(ws, r'^margen contribucion|^margen de contribucion',
                     obligatoria=False)
    f_be = _fila(ws, r'^break-?even \(meses\)|^break-?even$', fname=fname)
    f_cf = _fila(ws, r'^costes fijos mensuales', obligatoria=False)
    motor.f(ws, 'B' + str(f_be),
            "=IFERROR(MATCH(TRUE,INDEX('" + hoja_caja + "'!" + meses[0]
            + str(capa['acum']) + ':' + meses[-1] + str(capa['acum'])
            + ">0,0),0),\"No alcanzado\")", bold=True)
    base = base_bloque(ws, r'^umbral de ventas mensual')
    motor.val(ws, 'A' + str(base), 'Umbral de ventas mensual (€)')
    if f_cf and f_margen:
        motor.f(ws, 'B' + str(base),
                '=IFERROR(B' + str(f_cf) + '/(1-B' + str(
                    _fila(ws, r'^food cost', fname=fname)) + '),"")',
                fmt=FMT_EUR, bold=True)
    nota(ws, 'A' + str(base + 1),
         'El mes de break-even se lee sobre el flujo acumulado CON IVA y '
         'deuda de la hoja «' + hoja_caja + '». Si nunca cruza a positivo, '
         'dice «No alcanzado».')
    apunta(cambios, fname, ws, 'B' + str(f_be) + ' (la celda que da nombre al '
           'fichero) deja de estar vacía + umbral de ventas (TEC-03, DOM-09)')


# ==========================================================================
# §2.3.6 · calculadora-capex.xlsx — hoja de RANGOS de mercado
# ==========================================================================
RX_INSTR_RANGOS = re.compile(r'^Esta calculadora es la hoja de RANGOS')


def variante_capex(wb, fname=''):
    """`'rangos'` | `'desglosado'` | `'min-max'`, por la CABECERA."""
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        for fila in (3, 4):
            cab = [_norm(ws.cell(row=fila, column=c).value)
                   for c in range(1, ws.max_column + 1)]
            if any(c.startswith('rango bajo') for c in cab):
                return 'rangos', ws, fila
            if any(c.startswith('importe minimo') for c in cab):
                return 'min-max', ws, fila
            if 'partida' in cab and any(c.startswith('estimado') for c in cab):
                return 'desglosado', ws, fila
    raise VarianteDesconocida(
        fname + ': calculadora de CAPEX no reconocida. Hojas='
        + repr(wb.sheetnames))


def _fila_total(ws, fila_cab):
    """Fila TOTAL y última fila de datos por encima de ella, MEDIDAS."""
    f_total = None
    for r in range(fila_cab + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if v.startswith('total'):
                f_total = r
                break
        if f_total:
            break
    if f_total is None:
        return None, None
    ultima = f_total - 1
    while ultima > fila_cab and all(
            ws.cell(row=ultima, column=c).value is None
            for c in range(1, ws.max_column + 1)):
        ultima -= 1
    return f_total, ultima


def calculadora_capex(wb, fname, cambios, contenido, subtitulo):
    variante, ws, fila_cab = variante_capex(wb, fname)
    f_total, ultima = _fila_total(ws, fila_cab)
    if f_total is None:
        raise VarianteDesconocida(fname + ':' + ws.title
                                  + ': no encuentro la fila TOTAL')
    # Los SUM del total se rehacen SIEMPRE contra el rango MEDIDO. En panadería
    # `B31='=SUM(B4:B27)'` se dejaba fuera «Licencias + proyectos + tasas» y
    # «Fondo maniobra 4 meses»: 113.700 € cacheados donde la suma correcta es
    # 136.200 € (y 278.100 € frente a 316.100 € en el máximo).
    reparadas = []
    for c in range(1, ws.max_column + 1):
        cel = ws.cell(row=f_total, column=c)
        if cel.data_type != 'f' or 'SUM' not in str(cel.value).upper():
            continue
        from openpyxl.utils import get_column_letter as gcl
        col = gcl(c)
        rango = col + str(fila_cab + 1) + ':' + col + str(ultima)
        # RT-25 · la columna «Tu Presupuesto» llega VACÍA y su TOTAL imprimía
        # «0,00 €» mientras la celda de al lado, que sí guarda el caso «libro
        # en blanco», devolvía "". Misma incoherencia, misma columna.
        # `COUNT` (no `COUNTIF(...,"<>")`, ver RT-01) sólo cuenta números.
        propia = _norm(ws.cell(row=fila_cab, column=c).value).startswith(
            ('tu presupuesto', 'tu importe'))
        nueva = ('=IF(COUNT(' + rango + ')=0,"",SUM(' + rango + '))'
                 if propia else '=SUM(' + rango + ')')
        if _norm(cel.value) != _norm(nueva):
            reparadas.append(col + str(f_total) + ': ' + str(cel.value)
                             + ' → ' + nueva)
        motor.f(ws, col + str(f_total), nueva, fmt=FMT_EUR, bold=True)
    if reparadas:
        apunta(cambios, fname, ws, 'rango del TOTAL rehecho contra las filas '
               'MEDIDAS ' + str(fila_cab + 1) + '-' + str(ultima) + ': '
               + str(reparadas))

    # semáforo «tu presupuesto frente al rango», sólo donde hay rango
    if variante in ('rangos', 'min-max'):
        cols = _columnas_rango(ws, fila_cab)
        if cols:
            bajo, alto, propio = cols
            fila = f_total + 2
            motor.val(ws, 'A' + str(fila),
                      'Tu presupuesto frente al rango de mercado')
            motor.f(ws, propio + str(fila),
                    '=IF(' + propio + str(f_total) + '="","",IF(' + propio
                    + str(f_total) + '<' + bajo + str(f_total)
                    + ',"Por debajo del rango bajo",IF(' + propio
                    + str(f_total) + '>' + alto + str(f_total)
                    + ',"Por encima del rango alto","Dentro del rango")))')
            motor.fijar_formato(ws, propio + str(fila), 'General')
            apunta(cambios, fname, ws, 'aviso de encaje del presupuesto propio '
                   'en ' + propio + str(fila) + ' (§2.3.6)')
    _capex_rangos_y_notas(ws, fila_cab, ultima, fname, cambios, contenido)
    instruccion(wb, 'Esta calculadora es la hoja de RANGOS DE MERCADO. TU '
                    'CAPEX, el que va al banco, se rellena en '
                    'plan-financiero-3-anos.xlsx, hoja «Inversión», que trae '
                    'la correspondencia concepto → categoría de esta hoja. Los '
                    'dos libros se comparan; ninguno lee del otro (un .xlsx '
                    'movido de carpeta daría #REF!).', RX_INSTR_RANGOS)
    return variante


def _capex_rangos_y_notas(ws, fila_cab, ultima, fname, cambios, contenido):
    """RC-05 y RD-03 — el propio pack se desmentía en dos partidas.

    (a) El checklist de equipamiento de esta misma guía tasa 164.718,40 € y el
        «rango alto» de esta hoja decía 150.000 €.
    (b) El fondo de maniobra seguía siendo un rango tecleado de 60.000-200.000 €
        con la etiqueta «(6 meses)», cuando seis meses de la estructura que el
        P&L de este pack calcula son 935.510,40 €: el rango ALTO no cubre ni
        1,3 meses. El rango se queda (es mercado) y se dice al lado de dónde
        sale el número que vale.
    """
    conf = ((getattr(contenido, 'CAPEX', None) or {}) if contenido else {})
    if not conf:
        return 0
    from openpyxl.utils import get_column_letter as gcl
    cols = _columnas_rango(ws, fila_cab)
    # La columna se localiza por su CABECERA, no con `max_column + 1`: eso
    # añadiría una columna nueva en cada pasada (la trampa que ya documenta
    # `correspondencia_capex`, y que la 2.ª pasada delata como diferencias).
    titulo_nota = 'De dónde sale este rango'
    indice = None
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(row=fila_cab, column=c).value) == _norm(titulo_nota):
            indice = c
            break
    col_nota = gcl(indice or (ws.max_column + 1))
    tocadas = 0
    for patron, (bajo, medio, alto) in (conf.get('rangos') or {}).items():
        r = _fila(ws, patron, col=2, desde=fila_cab + 1, hasta=ultima,
                  obligatoria=False)
        if r is None or not cols:
            continue
        from openpyxl.utils import column_index_from_string as cifs
        c_bajo = cifs(cols[0])
        for off, valor in enumerate((bajo, medio, alto)):
            motor.val(ws, gcl(c_bajo + off) + str(r), valor, fmt=FMT_EUR)
        tocadas += 1
    notas = conf.get('notas') or {}
    if notas:
        cel = ws[col_nota + str(fila_cab)]
        cel.value = titulo_nota
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor='2D2D2D')
        cel.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col_nota].width = 52.0
        for patron, texto in notas.items():
            r = _fila(ws, patron, col=2, desde=fila_cab + 1, hasta=ultima,
                      obligatoria=False)
            if r is not None:
                nota(ws, col_nota + str(r), texto)
                tocadas += 1
    if tocadas:
        apunta(cambios, fname, ws, str(tocadas) + ' rangos/notas de CAPEX '
               'cuadrados con lo que este mismo pack tasa: el equipamiento '
               'sube a 180.000 € en el rango alto y el fondo de maniobra dice '
               'de dónde sale el número que vale (RC-05, RD-03)')
    return tocadas


def _columnas_rango(ws, fila_cab):
    from openpyxl.utils import get_column_letter as gcl
    bajo = alto = propio = None
    for c in range(2, ws.max_column + 1):
        cab = _norm(ws.cell(row=fila_cab, column=c).value)
        if cab.startswith('rango bajo') or cab.startswith('importe minimo'):
            bajo = gcl(c)
        elif cab.startswith('rango alto') or cab.startswith('importe maximo'):
            alto = gcl(c)
        elif cab.startswith('tu presupuesto') or cab.startswith('tu importe'):
            propio = gcl(c)
    return (bajo, alto, propio) if (bajo and alto and propio) else None


# ==========================================================================
# CONTRATO con main.py
# ==========================================================================
DESPACHO = {
    'calculadora-ticket-medio.xlsx': ticket_medio,
    'pl-mensual-escenarios.xlsx': pl_mensual,
    'plan-financiero-3-anos.xlsx': plan_financiero,
    'cash-flow-break-even.xlsx': cash_flow,
    'calculadora-capex.xlsx': calculadora_capex,
}
VARIANTES = {}


def _subtitulo(wb):
    for ws in wb.worksheets:
        v = ws['A2'].value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return 'AI Chef Pro · aichef.pro'


def post(wb, fname, cambios, registro, contenido):
    """Se ejecuta DESPUÉS de `motor.aplicar` y ANTES de `motor.cerrar`.

    Las fórmulas se escriben con `motor.f`, así que quedan registradas y
    `main.py` verifica una por una que acabaron con valor cacheado.
    """
    fn = DESPACHO.get(fname)
    if fn is None:
        return
    subtitulo = _subtitulo(wb)
    if fn is ticket_medio or fn is pl_mensual:
        variante = fn(wb, fname, cambios, contenido)
    else:
        variante = fn(wb, fname, cambios, contenido, subtitulo)
    VARIANTES[fname] = variante
    cambios.append(fname + ': grupo_a variante ' + repr(variante) + ' ('
                   + SPEC + ')')


# ==========================================================================
# DEMOSTRACIONES (pycel) — sobre COPIAS desechables, nunca sobre entregables
# ==========================================================================
def _compilar(carpeta, destino, fname):
    import os
    import shutil
    from pycel import ExcelCompiler
    os.makedirs(destino, exist_ok=True)
    copia = os.path.join(destino, 'a_' + fname)
    shutil.copy2(os.path.join(carpeta, fname), copia)
    return ExcelCompiler(filename=copia)


def _ev(xl, ref):
    import contextlib
    import os
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__ + ':' + str(e)[:90]


def _set(xl, ref, valor):
    _ev(xl, ref)
    xl.set_value(ref, valor)


def _mover(xl, salidas, entradas):
    """Evalúa `salidas`, cambia `entradas` y vuelve a evaluar `salidas`.

    ⚠️ El orden NO es cosmético y costó las cuatro primeras demostraciones de
    este módulo: `pycel` sólo propaga un `set_value` por el grafo de
    dependencias **que ya ha construido**, y el grafo se construye al evaluar
    una SALIDA. Cambiando primero el input, la salida se leía del caché del
    fichero y la demostración «pasaba» enseñando el valor viejo — que es la
    forma más cara de equivocarse: una demo que confirma lo que no ha probado.
    """
    antes = [_ev(xl, s) for s in salidas]
    for ref, valor in entradas:
        _set(xl, ref, valor)
    return antes, [_ev(xl, s) for s in salidas]


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _busca(carpeta, fname, hoja, patron):
    """Fila por etiqueta en el fichero YA escrito (las demos no adivinan)."""
    import os
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
    if hoja not in wb.sheetnames:
        return None, wb
    return _fila(wb[hoja], patron, obligatoria=False), wb


def _demo_constantes(carpeta, destino):
    """§1.2/§7-bis.12 — «conservar el número» con gate.

    Cada constante que el grupo convirtió en fórmula se RECALCULA con pycel y
    se compara con el valor que había. Coincidencia a 0,01 € o, si no coincide,
    la diferencia tiene que traer la nota que la justifica — que es el caso de
    `NUEVO-02`, donde el número viejo estaba mal sumado y corregirlo es
    precisamente el arreglo.
    """
    import os
    fuera, fallos = [], []
    for fname, registros in sorted(SUSTITUCIONES.items()):
        if not os.path.isfile(os.path.join(carpeta, fname)):
            continue
        xl = _compilar(carpeta, destino, fname)
        for reg in registros:
            ref = "'" + reg['hoja'] + "'!" + reg['celda']
            nuevo = _ev(xl, ref)
            iguala = (_num(nuevo)
                      and abs(nuevo - reg['anterior']) <= 0.01)
            fila = dict(reg, recalculado=nuevo, coincide=iguala)
            fuera.append(fila)
            if iguala or reg.get('nota'):
                continue
            fallos.append(fname + ':' + ref + ': la fórmula nueva da '
                          + repr(nuevo) + ' donde la constante decía '
                          + repr(reg['anterior']) + ' y NO hay nota que '
                          'justifique la diferencia (§1.2)')
    return {'sustituciones': fuera,
            'coinciden': sum(1 for x in fuera if x['coincide']),
            'diferencias_justificadas': sum(
                1 for x in fuera if not x['coincide'] and x.get('nota')),
            'fallos': fallos}


def demos(carpeta, destino, contenido):
    import os
    fuera, fallos = {}, []
    for nombre, fn in (('constantes_a_formulas',
                        lambda c, d: _demo_constantes(c, d)),
                       ('ticket_medio', _demo_ticket),
                       ('pl_escenarios', _demo_pl),
                       ('plan_financiero', _demo_plan),
                       ('cash_flow', _demo_cash),
                       ('capex', _demo_capex)):
        try:
            r = fn(carpeta, destino)
        except Exception as e:                               # noqa: BLE001
            # Una demo que revienta NO es una nota al pie: es una demostración
            # que no se ha hecho. La primera versión guardaba el error en el
            # informe y dejaba la tanda en verde — panadería pasó así con dos
            # demos caídas.
            r = {'error': type(e).__name__ + ': ' + str(e)[:300],
                 'fallos': ['demo ' + nombre + ' no se pudo ejecutar: '
                            + type(e).__name__ + ': ' + str(e)[:200]]}
        if r is None:
            continue
        fallos += r.pop('fallos', [])
        fuera[nombre] = r
    if not any(os.path.isfile(os.path.join(carpeta, f)) for f in FICHEROS):
        return {'grupo_a': {'aplica': False}, 'fallos': []}
    return {'grupo_a': fuera, 'fallos': fallos}


def _demo_ticket(carpeta, destino):
    """§2.1 — el simulador simula: mover un precio mueve el ticket, mover los
    días mueve la facturación mensual y el control del mix se pone en rojo si
    los tramos no reparten el 100 % de los comensales."""
    import os
    fname = 'calculadora-ticket-medio.xlsx'
    if not os.path.isfile(os.path.join(carpeta, fname)):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
    variante, fila_cab = variante_ticket(wb, fname)
    if variante == 'mix-producto':
        return _demo_ticket_mix(carpeta, destino, wb, fname, fila_cab)
    f_ticket = _fila(wb['Ticket Medio'], RX_TICKET_RES.pattern,
                     obligatoria=False)
    if f_ticket is None:
        return {'fallos': [fname + ': no hay fila de TICKET MEDIO tras el '
                           'grupo A']}
    ws = wb['Ticket Medio']
    f_mix = _fila(ws, RX_ETIQUETA_MIX.pattern, obligatoria=False)
    f_fmes = _fila(ws, r'^facturacion mensual', obligatoria=False)
    f_dias = _fila(ws, r'^dias abierto', obligatoria=False)
    pares, sueltos = pares_mix(ws, 4, f_ticket - 1)
    P = "'Ticket Medio'!"
    fallos = []
    r = {'fichero': fname, 'celda_ticket': 'B' + str(f_ticket),
         'pares_pct_precio': pares, 'incondicionales': sueltos}

    # (1) subir un PRECIO sube el ticket ponderado
    precio = 'B' + str(pares[0][1])
    xl = _compilar(carpeta, destino, fname)
    v0 = _ev(xl, P + precio)
    antes, despues = _mover(xl, [P + 'B' + str(f_ticket)],
                            [(P + precio, (v0 or 0) + 10)])
    r['precio_movido'] = {'celda': precio,
                          'etiqueta': ws['A' + str(pares[0][1])].value,
                          'de': v0, 'a': (v0 or 0) + 10}
    r['ticket'] = {'antes': antes[0], 'despues': despues[0]}
    if not (_num(antes[0]) and _num(despues[0]) and despues[0] > antes[0]):
        fallos.append(fname + ': subir ' + precio + ' 10 € no sube el ticket '
                      'ponderado (' + repr(antes[0]) + ' → '
                      + repr(despues[0]) + ')')

    # (2) los días abiertos sólo mueven la facturación MENSUAL
    if f_fmes and f_dias:
        xl = _compilar(carpeta, destino, fname)
        d0 = _ev(xl, P + 'B' + str(f_dias))
        antes, despues = _mover(
            xl, [P + 'B' + str(f_fmes), P + 'B' + str(f_ticket)],
            [(P + 'B' + str(f_dias), (d0 or 0) + 4)])
        r['dias_abiertos'] = {'de': d0, 'a': (d0 or 0) + 4}
        r['facturacion_mensual'] = {'antes': antes[0], 'despues': despues[0]}
        r['ticket_no_se_mueve_con_los_dias'] = (antes[1], despues[1])
        # Sin `contenido_<pid>/a.py` la hoja queda con las fórmulas y SIN
        # ejemplo: cubiertos/día y días son celdas verdes vacías y la
        # facturación vale 0 con razón. Exigir ahí que «suba» sería exigir que
        # el módulo de contenido exista, que es otro gate y de otra tanda.
        # ⚠️ La guarda mira la FACTURACIÓN, no el ticket: los 5 hermanos traen
        # el ticket precargado de fábrica pero NO los cubiertos/día, que son
        # filas que añade el grupo. Mirando el ticket, la guarda no saltaba y
        # los cinco daban un fallo que no era suyo.
        if not _num(antes[0]) or antes[0] == 0:
            r['sin_ejemplo_precargado'] = True
        elif not (_num(antes[0]) and _num(despues[0])
                  and despues[0] > antes[0]):
            fallos.append(fname + ': +4 días abiertos no suben la facturación '
                          'mensual (' + repr(antes[0]) + ' → '
                          + repr(despues[0]) + ')')

    # (3) el control del mix: con los tres tramos al 40 % debe dar 1,2
    if f_mix:
        formula = ws['B' + str(f_mix)].value or ''
        tramos = [p for p, _q in pares if ('B' + str(p)) in formula]
        xl = _compilar(carpeta, destino, fname)
        antes, despues = _mover(xl, [P + 'B' + str(f_mix)],
                                [(P + 'B' + str(p), 0.4) for p in tramos])
        r['control_del_mix'] = {'celda': 'B' + str(f_mix),
                                'tramos': tramos,
                                'con_el_ejemplo_precargado': antes[0],
                                'con_los_tres_al_40pct': despues[0]}
        if not (_num(antes[0]) and abs(antes[0] - 1) < 0.001):
            fallos.append(fname + ': el ejemplo precargado no reparte el 100 % '
                          'de los comensales (' + repr(antes[0]) + ')')
        if not (_num(despues[0]) and abs(despues[0] - 1.2) < 0.001):
            fallos.append(fname + ': con los tres tramos al 40 % el control '
                          'debería dar 1,2 y da ' + repr(despues[0]))
    return dict(r, fallos=fallos)


def _demo_ticket_mix(carpeta, destino, wb, fname, fila_cab):
    """Panadería: el ticket es la suma de `PVP × mix`, y el control es que el
    mix sume 100. Subir el PVP de un producto tiene que subir el ticket."""
    from openpyxl.utils import get_column_letter as gcl
    ws = wb['Ticket Medio']
    f_tot = _fila(ws, r'^ticket medio', fname=fname)
    col_mix = col_ap = col_pvp = None
    for c in range(2, ws.max_column + 1):
        cab = _norm(ws.cell(row=fila_cab, column=c).value)
        if 'mix' in cab:
            col_mix = gcl(c)
        elif 'aporte' in cab:
            col_ap = gcl(c)
        elif cab.startswith('pvp'):
            col_pvp = gcl(c)
    P = "'Ticket Medio'!"
    fallos, r = [], {'fichero': fname, 'variante': 'mix-producto',
                     'fila_total': f_tot}
    xl = _compilar(carpeta, destino, fname)
    r['mix_total'] = _ev(xl, P + col_mix + str(f_tot))
    if not (_num(r['mix_total']) and abs(r['mix_total'] - 100) < 0.001):
        fallos.append(fname + ': el mix de ventas suma ' + repr(r['mix_total'])
                      + ' y debe sumar 100')
    primera = fila_cab + 1
    v0 = _ev(xl, P + col_pvp + str(primera))
    antes, despues = _mover(xl, [P + col_ap + str(f_tot)],
                            [(P + col_pvp + str(primera), (v0 or 0) + 1)])
    r['ticket'] = {'antes': antes[0], 'despues': despues[0],
                   'pvp_movido': col_pvp + str(primera), 'de': v0}
    if not (_num(antes[0]) and _num(despues[0]) and despues[0] > antes[0]):
        fallos.append(fname + ': subir 1 € el PVP de '
                      + repr(ws['A' + str(primera)].value)
                      + ' no sube el ticket medio (' + repr(antes[0]) + ' → '
                      + repr(despues[0]) + ')')
    return dict(r, fallos=fallos)


def _demo_pl(carpeta, destino):
    """§2.2 — el P&L encadena: subir una entrada mueve facturación, food cost,
    margen y EBITDA; con los ingresos a 0 el margen devuelve `""`."""
    import os
    fname = 'pl-mensual-escenarios.xlsx'
    ruta = os.path.join(carpeta, fname)
    if not os.path.isfile(ruta):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(ruta)
    variante = motor.variante_pl(wb)
    hoja_nombre = ('Escenarios' if variante == 'escenarios-columnas'
                   else ('Pesimista' if variante == 'tres-hojas'
                         else 'P&L 3 escenarios'))
    ws = wb[hoja_nombre]
    P = "'" + hoja_nombre + "'!"
    fallos, r = [], {'fichero': fname, 'variante': variante,
                     'hoja': hoja_nombre}
    f_fact = _fila(ws, r'^facturacion mensual|^total ingresos'
                       r'|^facturacion total mensual', obligatoria=False)
    # ⚠️ El orden importa: `^food cost` casaría antes con la FILA DE
    # PARÁMETRO «Food cost (%)» del representante (fila 11) que con la de
    # resultado «Coste materia prima (€)» (fila 18), y la demostración
    # comprobaría que un porcentaje de entrada no cambia — que es cierto y no
    # demuestra nada.
    f_mp = (_fila(ws, r'^coste materia prima|^coste materias primas',
                  obligatoria=False)
            or _fila(ws, r'^food cost \(', obligatoria=False))
    f_eb = _fila(ws, RX_EBITDA, obligatoria=False)
    f_me = _fila(ws, RX_PCT_EBITDA, obligatoria=False)
    f_mov = _fila(ws, r'^cubiertos.*comida|^ventas sala|^tickets/dia',
                  obligatoria=False)
    if not (f_fact and f_eb):
        return {'fallos': [fname + ': no encuentro facturación o EBITDA tras '
                           'el grupo A'], 'variante': variante}
    salidas = [P + 'B' + str(x) for x in (f_fact, f_mp, f_eb, f_me) if x]
    nombres = [n for n, x in (('facturacion', f_fact), ('materia_prima', f_mp),
                              ('ebitda', f_eb), ('margen', f_me)) if x]
    if f_mov:
        xl = _compilar(carpeta, destino, fname)
        v0 = _ev(xl, P + 'B' + str(f_mov))
        if _num(v0):
            antes, despues = _mover(xl, salidas,
                                    [(P + 'B' + str(f_mov), v0 * 1.2)])
            r['entrada_movida'] = {'celda': 'B' + str(f_mov),
                                   'etiqueta': ws['A' + str(f_mov)].value,
                                   'de': v0, 'a': round(v0 * 1.2, 2)}
            r['cascada'] = dict((n, {'antes': a, 'despues': d})
                                for n, a, d in zip(nombres, antes, despues))
            for n, a, d in zip(nombres, antes, despues):
                if n == 'margen':
                    continue
                if not (_num(a) and _num(d) and d > a):
                    fallos.append(fname + ': subir «'
                                  + str(ws['A' + str(f_mov)].value)[:40]
                                  + '» un 20 % no mueve ' + n + ' ('
                                  + repr(a) + ' → ' + repr(d) + ')')
    # con los ingresos a 0 el margen devuelve "" (§7-bis.13), nunca «0,0 %»
    if f_me:
        xl = _compilar(carpeta, destino, fname)
        ceros = [(P + 'B' + str(fila), 0)
                 for fila in _lineas_de_ingreso(ws, f_fact)]
        _antes, despues = _mover(xl, [P + 'B' + str(f_me),
                                      P + 'B' + str(f_fact)], ceros)
        r['con_ingresos_a_cero'] = {'facturacion': despues[1],
                                    'margen': despues[0]}
        if despues[0] not in ('', None):
            fallos.append(fname + ': con la facturación a 0 el margen devuelve '
                          + repr(despues[0]) + ' en vez de "" (§7-bis.13)')
    return dict(r, fallos=fallos)


def _lineas_de_ingreso(ws, f_fact):
    """Celdas de entrada que alimentan la facturación (para ponerla a cero)."""
    f_ing = _fila(ws, RX_BLOQUE_ING, obligatoria=False)
    if f_ing and f_ing < f_fact:
        return [r for r in range(f_ing + 1, f_fact)
                if ws['B' + str(r)].data_type != 'f']
    return [r for r, _n, _c in etiquetas(ws, 1, 5, f_fact - 1)
            if ws['B' + str(r)].data_type != 'f']


def _demo_plan(carpeta, destino):
    """§2.3 — EBITDA sin amortización + EBIT, cuadro francés con su valor de
    control y sus dos guardas, y la proyección siguiendo al P&L."""
    import os
    fname = 'plan-financiero-3-anos.xlsx'
    ruta = os.path.join(carpeta, fname)
    if not os.path.isfile(ruta):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(ruta)
    fallos, r = [], {'fichero': fname, 'hojas': wb.sheetnames}

    # (1) TEC-08 con el escenario EXACTO que midió el R1: ventas 140.000 € y
    #     amortización 6.000 € → EBITDA 32.000 € y EBIT 26.000 €. La v1.1 daba
    #     26.000 € rotulado «EBITDA»: un 23 % por debajo del real.
    if 'P&L Mensual' in wb.sheetnames:
        ws = wb['P&L Mensual']
        P = "'P&L Mensual'!"
        f_ing_bloque = _fila(ws, RX_BLOQUE_ING, obligatoria=False)
        f_ing = _fila(ws, RX_TOT_INGRESOS, obligatoria=False)
        f_var = _fila(ws, RX_BLOQUE_VAR, obligatoria=False)
        f_tot_var = _fila(ws, RX_TOT_VARIABLES, obligatoria=False)
        f_fij = _fila(ws, RX_BLOQUE_FIJ, obligatoria=False)
        f_amort = _fila(ws, r'^amortizacion equipamiento', obligatoria=False)
        f_tot_fij = _fila(ws, RX_TOT_FIJOS, obligatoria=False)
        f_eb = _fila(ws, RX_EBITDA, obligatoria=False)
        f_me = _fila(ws, RX_PCT_EBITDA, obligatoria=False)
        f_ebit = _fila(ws, r'^ebit ', obligatoria=False)
        if f_ing and f_amort and f_eb and f_ebit:
            r['margen_formato'] = ws['B' + str(f_me)].number_format
            if r['margen_formato'] != FMT_PCT:
                fallos.append(fname + ": 'P&L Mensual'!B" + str(f_me)
                              + ' sigue en ' + repr(r['margen_formato'])
                              + ' y no en 0.0% (TEC-09)')
            r['C' + str(f_me) + '_vaciada'] = ws['C' + str(f_me)].value
            if ws['C' + str(f_me)].value is not None:
                fallos.append(fname + ": 'P&L Mensual'!C" + str(f_me)
                              + ' sigue dividiendo un porcentaje entre la '
                                'facturación (TEC-10)')
            xl = _compilar(carpeta, destino, fname)
            # El P&L ya NO se entrega en blanco (RD-01/RT-02/RC-01: llegaba con
            # «TOTAL INGRESOS 0,00 €» y «EBITDA 0,00 €»), así que el caso
            # «libro en blanco» hay que PROVOCARLO vaciando las entradas. Leer
            # las celdas tal cual mediría la precarga, no la guarda.
            todas = ([P + 'B' + str(x) for x in range(f_ing_bloque + 1, f_ing)]
                     + [P + 'B' + str(x) for x in range(f_var + 1, f_tot_var)]
                     + [P + 'B' + str(x) for x in range(f_fij + 1, f_tot_fij)])
            _mover(xl, [P + 'B' + str(f_me), P + 'C' + str(f_eb)],
                   [(ref, '') for ref in todas])
            r['con_el_libro_en_blanco'] = {
                'margen': _ev(xl, P + 'B' + str(f_me)),
                'pct_s_ventas_del_ebitda': _ev(xl, P + 'C' + str(f_eb))}
            for clave, valor in r['con_el_libro_en_blanco'].items():
                if valor not in ('', None, 0):
                    fallos.append(fname + ': con el libro en blanco «' + clave
                                  + '» dice ' + repr(valor)
                                  + ' (§7-bis.13: debe ser "")')
            # TODAS las líneas de ingreso a cero antes de poner los 140.000 €:
            # con la precarga hay cuatro, no una, y sumarlas daba 290.000 €.
            entradas = [(P + 'B' + str(x), 0)
                        for x in range(f_ing_bloque + 1, f_ing)]
            entradas += [(P + 'B' + str(f_ing_bloque + 1), 140000)]
            entradas += [(P + 'B' + str(x), 0)
                         for x in range(f_var + 1, f_tot_var)]
            entradas += [(P + 'B' + str(x), 0)
                         for x in range(f_fij + 1, f_tot_fij)]
            entradas += [(P + 'B' + str(f_fij + 1), 108000),
                         (P + 'B' + str(f_amort), 6000)]
            salidas = [P + 'B' + str(x) for x in
                       (f_ing, f_tot_fij, f_eb, f_ebit, f_me)]
            _antes, despues = _mover(xl, salidas, entradas)
            r['escenario_R1'] = dict(zip(
                ('ventas', 'costes_fijos_sin_amortizacion', 'ebitda', 'ebit',
                 'margen_ebitda'), despues))
            eb, ebit = despues[2], despues[3]
            if not (_num(eb) and abs(eb - 32000) < 0.01):
                fallos.append(fname + ': con ventas 140.000 €, fijos 108.000 € '
                              'y amortización 6.000 €, el EBITDA debe ser '
                              '32.000 € y es ' + repr(eb))
            if not (_num(ebit) and abs(ebit - 26000) < 0.01):
                fallos.append(fname + ': el EBIT debe ser 26.000 € (el EBITDA '
                              'menos la amortización) y es ' + repr(ebit))
            r['amortizacion_que_separa_ebitda_de_ebit'] = (
                round(eb - ebit, 2) if _num(eb) and _num(ebit) else None)

    # (1-bis) NUEVO-02 (§2.5): el «EBITDA» que valía exactamente la
    #         facturación (300.000 €) y el margen del 122,97 %.
    if 'P&L 3 años' in wb.sheetnames:
        ws = wb['P&L 3 años']
        P = "'P&L 3 años'!"
        f_fact = _fila(ws, r'^facturacion total', obligatoria=False)
        f_var = _fila(ws, RX_BLOQUE_VAR, obligatoria=False)
        f_tot = _fila(ws, r'^total costes', obligatoria=False)
        f_eb = _fila(ws, RX_EBITDA, desde=f_tot or 1, obligatoria=False)
        f_pct = _fila(ws, RX_PCT_EBITDA, desde=f_eb or 1, obligatoria=False)
        ultima = _ultimo_detalle(ws, f_var + 1, f_tot - 1, (RX_BLOQUE_FIJ,))
        xl = _compilar(carpeta, destino, fname)
        r['NUEVO_02'] = {}
        for col in ('B', 'C', 'D'):
            fact = _ev(xl, P + col + str(f_fact))
            tot = _ev(xl, P + col + str(f_tot))
            eb = _ev(xl, P + col + str(f_eb))
            pct = _ev(xl, P + col + str(f_pct))
            suma = sum(v for v in (_ev(xl, P + col + str(x))
                                   for x in range(f_var + 1, ultima + 1))
                       if _num(v))
            r['NUEVO_02'][col] = {'facturacion': fact, 'total_costes': tot,
                                  'suma_celda_a_celda': round(suma, 2),
                                  'ebitda': eb, 'margen': pct}
            if not (_num(tot) and abs(tot - suma) < 0.01):
                fallos.append(fname + ": 'P&L 3 años'!" + col + str(f_tot)
                              + ': el TOTAL COSTES vale ' + repr(tot)
                              + ' y la suma de sus filas ' + str(round(suma, 2))
                              + ' (NUEVO-02)')
            if not (_num(eb) and _num(fact) and _num(tot)
                    and abs(eb - (fact - tot)) < 0.01):
                fallos.append(fname + ": 'P&L 3 años'!" + col + str(f_eb)
                              + ': el EBITDA vale ' + repr(eb)
                              + ' y facturación − costes son '
                              + repr((fact - tot) if _num(fact) and _num(tot)
                                     else None) + ' (NUEVO-02)')
            if _num(eb) and _num(fact) and abs(eb - fact) < 0.01:
                fallos.append(fname + ": 'P&L 3 años'!" + col + str(f_eb)
                              + ': el EBITDA sigue valiendo EXACTAMENTE la '
                                'facturación (NUEVO-02)')
            if not (_num(pct) and 0 < pct < 0.5):
                fallos.append(fname + ": 'P&L 3 años'!" + col + str(f_pct)
                              + ': el margen EBITDA es ' + repr(pct)
                              + ', fuera de un rango creíble (la nota de la '
                                'hoja anuncia 15-22 %) (NUEVO-02)')

    # (2) cuadro francés: el valor de control ya verificado en la familia
    if HOJA_FIN in wb.sheetnames:
        F = "'" + HOJA_FIN + "'!"
        salidas = ([F + 'B12'] +
                   [F + c + str(18 + i) for i in range(ANOS_CUADRO)
                    for c in 'BCDF'])
        xl2 = _compilar(carpeta, destino, fname)
        _antes, d = _mover(xl2, salidas, [(F + 'B5', 100000), (F + 'B6', 5),
                                          (F + 'B7', 0.05), (F + 'B8', 0)])
        r['cuota_mensual_100k_5pct_60meses'] = d[0]
        if not (_num(d[0]) and abs(d[0] - 1887.12) < 0.01):
            fallos.append(fname + ': la anualidad de 100.000 € al 5 % en 60 '
                          'meses debe dar 1.887,12 €/mes y da ' + repr(d[0]))
        cuadro = [dict(zip(('capital_inicial', 'cuota', 'intereses',
                            'pendiente'), d[1 + i * 4:5 + i * 4]))
                  for i in range(ANOS_CUADRO)]
        r['cuadro_100k_5anos'] = cuadro
        pend = cuadro[4]['pendiente']
        if not (_num(pend) and abs(pend) < 0.01):
            fallos.append(fname + ': con plazo 5 el capital pendiente al final '
                          'del año 5 debe ser 0 y es ' + repr(pend))
        apagados = cuadro[5:]
        if any(not _num(x['cuota']) or abs(x['cuota']) > 0.01
               for x in apagados):
            fallos.append(fname + ': pasado el vencimiento el cuadro no se '
                          'apaga a 0 numérico: '
                          + repr([x['cuota'] for x in apagados]))
        # carencia ≥ plazo: aviso de TEXTO en C, y B, D y F NUMÉRICAS
        xl3 = _compilar(carpeta, destino, fname)
        _antes, g = _mover(xl3, [F + 'C18', F + 'B18', F + 'D18', F + 'F18'],
                           [(F + 'B5', 100000), (F + 'B6', 3),
                            (F + 'B7', 0.05), (F + 'B8', 3)])
        r['carencia_igual_al_plazo'] = dict(zip(('C18', 'B18', 'D18', 'F18'),
                                                g))
        if g[0] != AVISO_CARENCIA:
            fallos.append(fname + ': con carencia = plazo, C18 debe avisar y '
                          'dice ' + repr(g[0]))
        for etiqueta, valor in zip(('B18', 'D18', 'F18'), g[1:]):
            if not _num(valor):
                fallos.append(fname + ': con carencia = plazo, ' + etiqueta
                              + ' deja de ser numérica (' + repr(valor)
                              + '): propagaría #¡VALOR! al P&L y al cash flow')

    # (3) la proyección sigue al P&L y responde a los porcentajes
    if HOJA_PROY in wb.sheetnames and 'P&L Mensual' in wb.sheetnames:
        ws = wb['P&L Mensual']
        f_ing_bloque = _fila(ws, RX_BLOQUE_ING, obligatoria=False)
        Q = "'" + HOJA_PROY + "'!"
        f_tot_ing2 = _fila(ws, RX_TOT_INGRESOS, obligatoria=False)
        meses_pl, total_pl = columnas_mes(ws, 4)
        col_ing = (total_pl or 'B') + str(f_tot_ing2)
        xl4 = _compilar(carpeta, destino, fname)
        # El «libro en blanco» sólo se puede exigir si el P&L lo está: los 5
        # hermanos vienen con los doce meses precargados y ahí la proyección
        # DEBE dar un número. Comprobarlo sin mirar antes es pedirle a la hoja
        # que mienta.
        # B-01 · las filas de la cuenta de resultados bajaron 2 al entrar los
        # dos parámetros de la rampa (A9/A10). Se leen del propio libro para
        # que la demo no viva de un número copiado.
        wsp = wb[HOJA_PROY]
        f_ing_proy = _fila(wsp, r'^ingresos', desde=10, obligatoria=False) or 13
        f_neto = _fila(wsp, r'^resultado neto', obligatoria=False)
        f_meb = _fila(wsp, r'^margen ebitda', obligatoria=False)
        b_ing, d_ing = Q + 'B' + str(f_ing_proy), Q + 'D' + str(f_ing_proy)
        m_rampa, p_rampa = wsp['B9'].value, wsp['B10'].value
        r['rampa_ano_1'] = {'meses': m_rampa, 'primer_mes': p_rampa,
                            'fila_ingresos': f_ing_proy}
        r['pl_precargado'] = _ev(xl4, "'P&L Mensual'!" + col_ing)
        if not _num(r['pl_precargado']) or r['pl_precargado'] == 0:
            r['proyeccion_con_el_pl_en_blanco'] = {
                'ingresos_ano_1': _ev(xl4, b_ing),
                'resultado_neto_ano_3': _ev(xl4, Q + 'D' + str(f_neto))
                if f_neto else None,
                'margen_ebitda_ano_3': _ev(xl4, Q + 'D' + str(f_meb))
                if f_meb else None}
            for clave, valor in r['proyeccion_con_el_pl_en_blanco'].items():
                if valor not in ('', None):
                    fallos.append(fname + ': con el P&L en blanco la proyección '
                                  'devuelve ' + repr(valor) + ' en ' + clave
                                  + ' (§7-bis.13: debe ser "")')
        _antes, d = _mover(xl4, [b_ing, d_ing,
                                 "'P&L Mensual'!" + col_ing],
                           [("'P&L Mensual'!B" + str(f_ing_bloque + 1),
                             140000)])
        r['proyeccion_ano_1'] = d[0]
        r['proyeccion_ano_3_base'] = d[1]
        r['ingresos_anuales_del_pl'] = d[2]
        crucero = (d[2] if total_pl else (d[2] * 12 if _num(d[2]) else None))
        factor = 12.0
        if _num(m_rampa) and m_rampa > 1 and _num(p_rampa):
            factor = m_rampa * (1 + p_rampa) / 2 + (12 - m_rampa)
        esperado = crucero * factor / 12 if _num(crucero) else None
        r['suma_factores_rampa'] = factor
        r['ano_1_esperado'] = esperado
        if not (_num(d[0]) and _num(esperado) and abs(d[0] - esperado) < 1):
            fallos.append(fname + ': el Año 1 de la proyección debe ser el mes '
                          'de crucero por la suma de los factores de la rampa '
                          '(' + repr(esperado) + ') y es ' + repr(d[0]))
        # B-01 · la rampa MANDA sobre el año 1, y con 0 meses la proyección
        # vuelve exactamente al «P&L × 12» de la v1.
        _antes, d0 = _mover(xl4, [b_ing], [(Q + 'B9', 0)])
        r['ano_1_con_rampa_0_meses'] = d0[0]
        if not (_num(d0[0]) and _num(crucero) and abs(d0[0] - crucero) < 1):
            fallos.append(fname + ': con 0 meses de rampa el Año 1 debe volver '
                          'al P&L anualizado (' + repr(crucero) + ') y da '
                          + repr(d0[0]))
        _antes, d12 = _mover(xl4, [b_ing], [(Q + 'B9', 12)])
        r['ano_1_con_rampa_12_meses'] = d12[0]
        if not (_num(d12[0]) and _num(d0[0]) and d12[0] < d0[0]):
            fallos.append(fname + ': alargar la rampa a 12 meses no baja el '
                          'Año 1 (' + repr(d0[0]) + ' → ' + repr(d12[0])
                          + '): las dos celdas verdes no gobiernan nada')
        _mover(xl4, [b_ing], [(Q + 'B9', m_rampa)])
        c0 = _ev(xl4, Q + 'C6')
        salidas3 = [d_ing] + ([Q + 'D' + str(f_neto)] if f_neto else [])
        _antes, d2 = _mover(xl4, salidas3, [(Q + 'C6', (c0 or 0) + 0.1)])
        r['crecimiento_ano_2'] = {'de': c0, 'a': (c0 or 0) + 0.1}
        r['proyeccion_ano_3_con_mas_crecimiento'] = d2[0]
        if _num(d[1]) and _num(d2[0]) and d2[0] <= d[1]:
            fallos.append(fname + ': +10 puntos de crecimiento en el Año 2 no '
                          'suben los ingresos del Año 3 (' + repr(d[1])
                          + ' → ' + repr(d2[0]) + ')')
    return dict(r, fallos=fallos)


def _demo_cash(carpeta, destino):
    """§2.4 — el acumulado encadena, el break-even sale en el mes que cruza a
    positivo y devuelve «No alcanzado» (no `#N/A`) si nunca cruza."""
    import os
    fname = 'cash-flow-break-even.xlsx'
    ruta = os.path.join(carpeta, fname)
    if not os.path.isfile(ruta):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(ruta)
    variante, nombre_caja, nombre_be = variante_cash(wb, fname)
    if variante == 'cash-plurianual':
        return _demo_cash_plurianual(carpeta, destino, wb, fname, nombre_caja)
    ws = wb[nombre_caja]
    fila_cab = motor.fila_cabecera_tabla(ws) or 4
    meses, _total = columnas_mes(ws, fila_cab)
    meses = meses[:12]
    f_ing_bloque = _fila(ws, r'^ingresos$|^entradas$', obligatoria=False)
    f_tot_ing = _fila(ws, r'^total ingresos$|^total entradas$',
                      obligatoria=False)
    f_gas_bloque = _fila(ws, r'^gastos$|^salidas$', obligatoria=False)
    f_tot_gas = _fila(ws, r'^total gastos$|^total salidas$', obligatoria=False)
    f_acum = _fila(ws, r'^flujo acumulado$|^saldo acumulado$',
                   obligatoria=False)
    f_acum_iva = _fila(ws, r'^flujo acumulado con iva', obligatoria=False)
    ws_be = wb[nombre_be] if nombre_be else ws
    f_be = _fila(ws_be, r'^mes de break-?even|^break-?even \(meses\)',
                 obligatoria=False)
    f_umbral = _fila(ws_be, r'^umbral de ventas', obligatoria=False)
    f_mc = _fila(ws_be, r'^margen de contribucion', obligatoria=False)
    P = "'" + nombre_caja + "'!"
    B = "'" + ws_be.title + "'!"
    fallos, r = [], {'fichero': fname, 'variante': variante, 'meses': meses,
                     'fila_acumulado': f_acum,
                     'fila_acumulado_con_iva_y_deuda': f_acum_iva}
    entradas_filas = [x for x in range(f_ing_bloque + 1, f_tot_ing)
                      if ws[meses[0] + str(x)].data_type != 'f']
    salidas_filas = [x for x in range(f_gas_bloque + 1, f_tot_gas)
                     if ws[meses[0] + str(x)].data_type != 'f']

    # serie creciente: ingresos 10.000 × mes contra un gasto fijo de 25.000
    entradas = []
    for i, col in enumerate(meses):
        entradas.append((P + col + str(entradas_filas[0]), 10000 * (i + 1)))
        entradas += [(P + col + str(x), 0) for x in entradas_filas[1:]]
        entradas.append((P + col + str(salidas_filas[0]), 25000))
        entradas += [(P + col + str(x), 0) for x in salidas_filas[1:]]
        if f_acum_iva:
            entradas.append((P + col + str(f_acum_iva - 4), 0))   # IVA soport.
            entradas.append((P + col + str(f_acum_iva - 2), 0))   # cuota
    salidas = ([P + c + str(f_acum) for c in meses]
               + ([P + c + str(f_acum_iva) for c in meses]
                  if f_acum_iva else [])
               + ([B + 'B' + str(f_be)] if f_be else []))
    xl = _compilar(carpeta, destino, fname)
    _antes, d = _mover(xl, salidas, entradas)
    r['acumulado'] = d[:12]
    r['acumulado_con_iva_y_deuda'] = d[12:24] if f_acum_iva else None
    serie = r['acumulado']
    if all(_num(x) for x in serie):
        neto = [10000 * (i + 1) - 25000 for i in range(12)]
        esperado, acumulado = [], 0.0
        for x in neto:
            acumulado += x
            esperado.append(acumulado)
        r['acumulado_esperado'] = esperado
        r['encadena'] = all(abs(a - b) < 0.01
                            for a, b in zip(serie, esperado))
        if not r['encadena']:
            fallos.append(fname + ': el flujo acumulado no encadena mes a mes. '
                          'calculado=' + repr(serie) + ' esperado='
                          + repr(esperado))
    else:
        fallos.append(fname + ': el flujo acumulado no evalúa: ' + repr(serie))
    if f_be:
        r['mes_de_break_even'] = d[-1]
        if not _num(r['mes_de_break_even']):
            fallos.append(fname + ': con una serie que cruza a positivo, el '
                          'mes de break-even devuelve '
                          + repr(r['mes_de_break_even'])
                          + ' en vez de un número')
        # serie SIEMPRE negativa → «No alcanzado», nunca #N/A
        xl2 = _compilar(carpeta, destino, fname)
        negativas = []
        for col in meses:
            negativas += [(P + col + str(x), 0) for x in entradas_filas]
            negativas.append((P + col + str(salidas_filas[0]), 25000))
            negativas += [(P + col + str(x), 0) for x in salidas_filas[1:]]
        _a, d2 = _mover(xl2, [B + 'B' + str(f_be)], negativas)
        r['mes_de_break_even_con_serie_siempre_negativa'] = d2[0]
        if d2[0] != 'No alcanzado':
            fallos.append(fname + ': con una serie siempre negativa el mes de '
                          'break-even devuelve ' + repr(d2[0])
                          + ' en vez de "No alcanzado"')
    if f_umbral and f_mc:
        xl3 = _compilar(carpeta, destino, fname)
        _a, d3 = _mover(xl3, [B + 'B' + str(f_umbral)],
                        [(B + 'B' + str(f_mc), 0)])
        r['umbral_con_margen_de_contribucion_cero'] = d3[0]
        if d3[0] not in ('', None):
            fallos.append(fname + ': con margen de contribución 0 el umbral '
                          'devuelve ' + repr(d3[0]) + ' en vez de "" '
                          '(debería atraparlo IFERROR)')
    return dict(r, fallos=fallos)


def _demo_cash_plurianual(carpeta, destino, wb, fname, nombre_caja):
    """Panadería: el break-even deja de ser el texto «M10-M12 (realista)».

    Se comprueba que la celda devuelve el mes en que la caja acumulada cruza a
    positivo y que, con una serie que nunca cruza, dice «No alcanzado» y no
    `#N/A`.
    """
    ws = wb[nombre_caja]
    fila_cab = motor.fila_cabecera_tabla(ws) or 3
    meses, _total = columnas_mes(ws, fila_cab)
    meses = meses[:12]
    f_acum = _fila(ws, r'^caja acumulada|^saldo acumulado|^flujo acumulado',
                   fname=fname)
    f_be = _fila(ws, r'^break-?even|^mes de break-?even', fname=fname)
    f_ing = _fila(ws, r'^ingresos$', fname=fname)
    f_capex = _fila(ws, r'^capex inicial', obligatoria=False)
    P = "'" + nombre_caja + "'!"
    fallos, r = [], {'fichero': fname, 'variante': 'cash-plurianual',
                     'fila_caja_acumulada': f_acum, 'fila_break_even': f_be}
    xl = _compilar(carpeta, destino, fname)
    r['caja_acumulada_con_el_ejemplo'] = [_ev(xl, P + c + str(f_acum))
                                          for c in meses]
    r['mes_de_break_even'] = _ev(xl, P + 'B' + str(f_be))
    serie = r['caja_acumulada_con_el_ejemplo']
    if all(_num(x) for x in serie) and any(x > 0 for x in serie):
        esperado = next(i + 1 for i, x in enumerate(serie) if x > 0)
        if r['mes_de_break_even'] != esperado:
            fallos.append(fname + ': la caja cruza a positivo en el mes '
                          + str(esperado) + ' y la celda dice '
                          + repr(r['mes_de_break_even']))
    elif r['mes_de_break_even'] != 'No alcanzado':
        fallos.append(fname + ': la caja no cruza a positivo en 12 meses y la '
                      'celda dice ' + repr(r['mes_de_break_even'])
                      + ' en vez de "No alcanzado"')
    # serie que SÍ cruza: ingresos crecientes contra un CAPEX pequeño
    xl2 = _compilar(carpeta, destino, fname)
    entradas = [(P + c + str(f_ing), 30000 + 3000 * i)
                for i, c in enumerate(meses)]
    if f_capex:
        entradas.append((P + 'B' + str(f_capex), -10000))
    _a, d = _mover(xl2, [P + 'B' + str(f_be)], entradas)
    r['mes_de_break_even_con_serie_que_cruza'] = d[0]
    if not _num(d[0]):
        fallos.append(fname + ': con una serie que cruza a positivo el mes de '
                      'break-even devuelve ' + repr(d[0]))
    return dict(r, fallos=fallos)


def _demo_capex(carpeta, destino):
    import os
    fname = 'calculadora-capex.xlsx'
    ruta = os.path.join(carpeta, fname)
    if not os.path.isfile(ruta):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(ruta)
    variante, ws, fila_cab = variante_capex(wb, fname)
    f_total, ultima = _fila_total(ws, fila_cab)
    xl = _compilar(carpeta, destino, fname)
    P = "'" + ws.title + "'!"
    fallos, r = [], {'fichero': fname, 'variante': variante,
                     'fila_total': f_total, 'ultima_fila_de_datos': ultima}
    from openpyxl.utils import get_column_letter as gcl
    r['totales'] = {}
    for c in range(2, ws.max_column + 1):
        cel = ws.cell(row=f_total, column=c)
        if cel.data_type != 'f':
            continue
        col = gcl(c)
        total = _ev(xl, P + col + str(f_total))
        suma = 0.0
        for fila in range(fila_cab + 1, ultima + 1):
            v = _ev(xl, P + col + str(fila))
            if _num(v):
                suma += v
        r['totales'][col + str(f_total)] = {'formula': cel.value,
                                            'valor': total,
                                            'suma_celda_a_celda': round(suma,
                                                                        2)}
        if _num(total) and abs(total - suma) > 0.01:
            fallos.append(fname + ':' + ws.title + '!' + col + str(f_total)
                          + ': el TOTAL vale ' + str(total) + ' y la suma de '
                          'sus filas ' + str(round(suma, 2)))
    return dict(r, fallos=fallos)
