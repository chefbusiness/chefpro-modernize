#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 de la familia «Guías Cómo Montar».

    python3 main.py --producto <pid> --dry-run [--solo motor|a|b|c] \
                    [--json informe.json]

REGLA DURA: `astro-site/public/dl/guia-*/` NO se toca. `--dry-run` regenera una
copia en el scratchpad y trabaja allí. Sin `--dry-run` el script **ABORTA**
salvo que se le pase `GUIAS_APPLY=1` en el entorno: la ejecución real la hace el
orquestador cuando T5 firme (SPEC §9). En producción se hace **respaldo previo
al scratchpad** antes de escribir in place — nunca un `.bak` dentro de
`public/dl/`, que se publicaría con el sitio.

Orden del pipeline (y el orden importa):
  1. Copia de trabajo (o respaldo previo).
  2. **Pre-vuelo de moldes** (§1.1/§7-bis.11): se abren los xlsx del producto y
     se detecta el molde de cada checklist ANTES de escribir un solo byte. Si
     uno no encaja, el script aborta con exit 2, deja el informe escrito y NO
     modifica nada. «Aplicar el molde A por defecto» rompería panadería y
     duplicaría el total de dark-kitchen.
  3. Por fichero: `grupo.pre` → `motor.aplicar` (§1) → `grupo.post` →
     `motor.cerrar` (§1: formatos, checklists, DV, verdes, protección, bio +
     versión) → `wb.save()`.
  4. Idempotencia: el pipeline entero se repite sobre un CLON del resultado y se
     compara celda a celda. Debe dar 0 diferencias.
  5. `inject_cache.py` sobre los ficheros tocados — SIEMPRE al final: cualquier
     `wb.save()` posterior borraría el caché que acaba de escribir.
  6. Verificación `data_only` de todas las fórmulas nuevas del registro.
  7. `censo-entregables.py --only <carpeta> --fail --quiet`.
  8. Demostraciones con pycel (el `% completado` responde al estado terminal
     REAL de cada lista, el TOTAL suma la columna de coste, el semáforo
     `ISNUMBER` no se enciende con un texto, la protección no lleva contraseña,
     y bio + versión aparecen una sola vez por fichero).

**Los módulos de grupo pueden no existir todavía.** `--solo motor` no carga
ninguno y el pipeline funciona igual. Un grupo que se PIDE y no carga es ROJO
(la v2.0 del kit anterior imprimía «se omite» y seguía hasta «TODO VERDE»: un
paquete a medio construir se declaraba terminado).

CONTRATO de `grupo_a.py` / `grupo_b.py` / `grupo_c.py` (en esta misma carpeta):

    FICHEROS = ['pl-mensual-escenarios.xlsx', ...]   # o `def ficheros(ctx)`
    PROPIOS  = [...]      # ficheros donde el §1 del motor NO se aplica (el
                          # grupo se encarga entero, incluido motor.cerrar)
    def pre(wb, fname, cambios, contenido): ...     # antes de motor.aplicar
    def post(wb, fname, cambios, registro, contenido): ...  # tras aplicar
    def demos(carpeta, origen, destino, contenido) -> dict  # su clave 'fallos'
                                                            # entra al veredicto

Los parámetros `contenido` y los que sobren son OPCIONALES: `_llamar()` inspecta
la firma y pasa sólo los que el grupo declare, para que un grupo escrito con la
firma corta del kit-plan-financiero siga funcionando aquí.

CONTENIDO POR PRODUCTO: `contenido_<pid con guiones bajos>/a.py|b.py|c.py`
(paquete o carpeta suelta). Cada submódulo expone el contenido de SU grupo para
ESA guía como diccionarios del módulo; `main.py` lo importa y se lo pasa al
grupo correspondiente. Si falta, el grupo recibe `None` y debe limitarse a lo
que pueda hacer sin contenido propio.

Las FÓRMULAS se escriben con `motor.f(ws, coord, formula, fmt)` —queda
registrada y aquí se verifica una por una que acabó con valor cacheado— o, si se
escriben a mano, se añaden a la lista `registro` como `(hoja, coord, formula)`.

Los grupos NO tienen que ocuparse de: protección, la hoja de Instrucciones, la
bio/versión, el nombre de la pestaña, el formato por tipo de dato, la DV de las
verdes ni el bloque de resumen de los checklists.

Térmica: todo en SERIE, un python cada vez. Sin builds ni navegador.
"""
import argparse
import contextlib
import datetime
import glob
import importlib
import importlib.util
import inspect
import json
import logging
import os
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    '9a6ebdb7-5d45-4ab1-9e93-10b86cb95c42/scratchpad/guias')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')
SPEC = 'scripts/productos-digitales/guias-v2-SPEC.md §1'

PRODUCTOS = ('guia-restaurante-gastronomico', 'guia-restaurante-casual',
             'guia-restaurante-mexicano', 'guia-restaurante-peruano',
             'guia-restaurante-japones', 'guia-restaurante-nikkei',
             'guia-panaderia-obrador', 'guia-dark-kitchen')


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Preparación
# ==========================================================================
def ficheros_de(carpeta):
    return sorted(os.path.basename(p)
                  for p in glob.glob(os.path.join(carpeta, '*.xlsx')))


def preparar_copia(origen, destino):
    if not os.path.isdir(origen):
        raise SystemExit('No existe el origen: ' + origen)
    if os.path.isdir(destino):
        shutil.rmtree(destino)
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    shutil.copytree(origen, destino)
    log('  copia de trabajo regenerada: ' + destino)


def digest(path):
    """Huella comparable de un .xlsx: valores, formato de número, relleno,
    bloqueo, merges, DV, formato condicional, protección, área de impresión y
    pie. Es lo que compara la idempotencia."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                if c.value is None and relleno is None:
                    continue
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked))
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted(str(dv.type) + ':' + str(dv.formula1) + ':'
                         + str(dv.sqref)
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted(str(cf.sqref) + ':' + str(len(cf.rules))
                         for cf in ws.conditional_formatting),
            'prot': bool(ws.protection.sheet),
            'anchos': dict((k, v.width) for k, v in
                           ws.column_dimensions.items() if v.width),
            'area': str(ws.print_area),
            'pie': str(ws.oddFooter.center.text),
        }
    fuera['__props__'] = {'subject': wb.properties.subject,
                          'title': wb.properties.title,
                          'creator': wb.properties.creator}
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append(fichero + ':' + hoja + ': hoja sólo en una pasada')
            continue
        if hoja == '__props__':
            if a[hoja] != b[hoja]:
                fuera.append(fichero + ': cambian las propiedades '
                             + str(a[hoja]) + ' → ' + str(b[hoja]))
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'prot', 'area', 'pie', 'anchos'):
            if ha[k] != hb[k]:
                fuera.append(fichero + ':' + hoja + ': cambia ' + k + ' ('
                             + str(ha[k])[:140] + ' → ' + str(hb[k])[:140]
                             + ')')
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append(fichero + ':' + hoja + '!' + coord + ': '
                             + str(ca.get(coord)) + ' → ' + str(cb.get(coord)))
    return fuera


# ==========================================================================
# Grupos y contenido
# ==========================================================================
def _llamar(fn, **kwargs):
    """Llama `fn` pasándole sólo los argumentos que su firma declara.

    Así un `post(wb, fname, cambios, registro)` escrito con la firma del
    kit-plan-financiero sigue funcionando aunque aquí se le ofrezca además
    `contenido`; y un grupo que sí lo necesite lo declara y lo recibe.
    """
    try:
        params = inspect.signature(fn).parameters
    except (TypeError, ValueError):                          # noqa: BLE001
        return fn(**kwargs)
    if any(p.kind == inspect.Parameter.VAR_KEYWORD
           for p in params.values()):
        return fn(**kwargs)
    orden = [n for n in params if n in kwargs]
    return fn(*[kwargs[n] for n in orden])


def cargar_contenido(pid, letra, avisos):
    """`contenido_<pid_>/a.py` — el contenido de ESE grupo para ESA guía."""
    paquete = 'contenido_' + pid.replace('-', '_')
    ruta = os.path.join(AQUI, paquete, letra + '.py')
    if not os.path.isfile(ruta):
        avisos.append(paquete + '/' + letra + '.py no existe — grupo_' + letra
                      + ' trabaja sin contenido propio')
        return None
    spec = importlib.util.spec_from_file_location(paquete + '.' + letra, ruta)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def ficheros_del_grupo(g, ctx):
    fn = getattr(g, 'ficheros', None)
    if callable(fn):
        return list(_llamar(fn, ctx=ctx))
    return list(getattr(g, 'FICHEROS', []))


# ==========================================================================
# Pre-vuelo de moldes (§1.1 / §7-bis.11)
# ==========================================================================
def prevuelo(carpeta, nombres, pid):
    """Detecta el molde de cada checklist ANTES de escribir. Aborta si no lo
    reconoce: nunca «molde A por defecto»."""
    detectado, fallos = [], []
    for fname in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        try:
            info = motor.detectar(wb, fname, pid)
        except motor.MoldeDesconocido as e:                  # noqa: BLE001
            fallos.append(str(e))
            continue
        detectado.append({'fichero': fname,
                          'variante_pl': info['variante_pl'],
                          'checklists': info['checklists']})
    return detectado, fallos


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, grupos, contenidos, informe_global, pid):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    cambios, registro_grupo = [], []
    motor.REGISTRO = []

    propio = any(fname in getattr(g, 'PROPIOS', []) for g, _ in grupos)

    for g, letra in grupos:
        if fname in ficheros_del_grupo(g, {'producto': pid}) \
                and hasattr(g, 'pre'):
            _llamar(g.pre, wb=wb, fname=fname, cambios=cambios,
                    contenido=contenidos.get(letra))

    if not propio:
        motor.aplicar(wb, fname, cambios, pid)

    for g, letra in grupos:
        if fname in ficheros_del_grupo(g, {'producto': pid}) \
                and hasattr(g, 'post'):
            _llamar(g.post, wb=wb, fname=fname, cambios=cambios,
                    registro=registro_grupo, contenido=contenidos.get(letra))

    info = {}
    if not propio:
        info = motor.cerrar(wb, fname, cambios, pid)

    literales = motor.literales_sospechosos(wb, fname)
    cuenta = motor.contadores(wb, fname)
    wb.save(path)

    registro = list(motor.REGISTRO)
    registro += [(h, c, fm) for h, c, fm in registro_grupo
                 if isinstance(fm, str) and fm.startswith('=')]
    if informe_global is not None:
        informe_global.append({
            'fichero': fname,
            'cambios': cambios,
            'formulas_nuevas': len(registro),
            'checklists': (info or {}).get('checklists_cerrados', []),
            'variante_pl': (info or {}).get('variante_pl'),
            # Los literales van ENTEROS al total y con muestra amplia: un
            # `slice(0, 12)` por fichero enseñó una vez sólo los del primer mes
            # y el lector creyó el problema acotado.
            'literales_sospechosos_total': len(literales),
            'literales_sospechosos': literales[:60],
            'contadores': cuenta})
    return registro


def ficheros_a_tocar(nombres, grupos, pid):
    fuera = list(nombres)
    for g, _ in grupos:
        for fn in ficheros_del_grupo(g, {'producto': pid}):
            if fn not in fuera:
                fuera.append(fn)
    return fuera


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        salida = r.stdout.strip()
        fuera.append((n, salida, r.returncode))
        log('    ' + salida)
        if r.returncode != 0:
            log('    ERROR inject_cache: ' + r.stderr[-400:])
    return fuera


def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA debe tener valor cacheado.

    Una celda vacía cuya fórmula contenga `""` NO se da por buena sin más: se
    EVALÚA con pycel, y si la rama que gana no es la alterna vacía, es un fallo
    de caché de verdad (el fichero se entregaría en blanco en Vista previa y en
    Google Sheets, que es el incidente que motivó `inject_cache.py`).
    """
    fallos, vacias, ok = [], 0, 0
    sospechosas, no_verificadas = {}, []
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wbv = openpyxl.load_workbook(path, data_only=True)
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append(fname + ':' + hoja + '!' + coord
                              + ': hoja ausente (¿renombrada después?)')
                continue
            v = wbv[hoja][coord].value
            if v is None:
                if '""' in formula:
                    vacias += 1
                    sospechosas.setdefault(fname, []).append((hoja, coord,
                                                              formula))
                else:
                    fallos.append(fname + ':' + hoja + '!' + coord
                                  + ': sin cache (' + formula[:60] + ')')
            else:
                ok += 1
    comprobadas = 0
    for fname, celdas in sospechosas.items():
        try:
            xl = _pycel(os.path.join(carpeta, fname))
        except Exception as e:                               # noqa: BLE001
            no_verificadas.append(fname + ': no se pudo compilar ('
                                  + type(e).__name__ + ')')
            continue
        for hoja, coord, formula in celdas:
            ref = "'" + hoja + "'!" + coord
            valor = _ev(xl, ref)
            comprobadas += 1
            if isinstance(valor, str) and valor.startswith('ERR:'):
                no_verificadas.append(fname + ':' + ref + ': ' + valor)
            elif valor not in ('', None):
                fallos.append(fname + ':' + ref + ': la fórmula devuelve '
                              + repr(valor) + ' pero se entregó SIN caché ('
                              + formula[:60] + ')')
    return {'con_valor': ok, 'vacias_por_diseno': vacias,
            'vacias_comprobadas_con_pycel': comprobadas,
            'vacias_no_verificadas': no_verificadas, 'fallos': fallos}


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-2000:])
    lineas = (r.stdout.strip().splitlines()
              + r.stderr.strip().splitlines())
    return {'exit': r.returncode, 'salida': lineas[-8:]}


# ==========================================================================
# Demostraciones (pycel) — sobre COPIAS desechables, nunca sobre entregables
# ==========================================================================
def _copia_demo(carpeta, fname, demos_dir):
    os.makedirs(demos_dir, exist_ok=True)
    destino = os.path.join(demos_dir, fname)
    shutil.copy2(os.path.join(carpeta, fname), destino)
    return destino


def demo_checklists(carpeta, nombres, demos_dir):
    """§1.9 — el bloque de resumen de cada checklist tiene que CALCULAR:

    · el TOTAL suma la columna de coste (o el que ya traía el molde C/D);
    · el `% completado` responde al estado terminal REAL de esa lista
      («Instalado» en `checklist-equipamiento-obra`, no «Completado»), y
      CAMBIA al marcar un ítem;
    · los subtotales por categoría suman lo mismo que el TOTAL.
    """
    fuera, fallos = [], []
    for fname in nombres:
        if 'checklist' not in fname.lower():
            continue
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            if ws.title == 'Instrucciones':
                continue
            fila_marca = coord_total = coord_avance = None
            subtot = []
            en_subtot = False
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str) and motor.RX_MARCA.match(v):
                    fila_marca = r
                    continue
                if fila_marca is None:
                    continue
                if v == motor.ETIQUETA_TOTAL:
                    for c in range(2, ws.max_column + 1):
                        if ws.cell(row=r, column=c).data_type == 'f':
                            coord_total = ws.cell(row=r,
                                                  column=c).coordinate
                elif v == motor.ETIQUETA_AVANCE:
                    for c in range(2, ws.max_column + 1):
                        if ws.cell(row=r, column=c).data_type == 'f':
                            coord_avance = ws.cell(row=r,
                                                   column=c).coordinate
                elif v == motor.ETIQUETA_SUBTOT:
                    en_subtot = True
                elif en_subtot and isinstance(v, str):
                    for c in range(2, ws.max_column + 1):
                        if ws.cell(row=r, column=c).data_type == 'f':
                            subtot.append(ws.cell(row=r,
                                                  column=c).coordinate)
            if fila_marca is None:
                # Molde B (panadería) no lleva bloque: su contador ya existe y
                # no tiene columna de coste (§3.3/§7-bis.17). Se comprueba que
                # ese contador SIGUE ahí, que es lo que se prometió respetar.
                contador = sum(1 for row in ws.iter_rows() for c in row
                               if c.data_type == 'f'
                               and 'COUNTIF' in str(c.value).upper())
                if contador:
                    fuera.append({'fichero': fname, 'hoja': ws.title,
                                  'molde_sin_bloque': True,
                                  'contador_propio_intacto': contador})
                else:
                    fallos.append(fname + ':' + ws.title + ': ni bloque de '
                                  'resumen (§1.9) ni contador propio')
                continue
            copia = _copia_demo(carpeta, fname, demos_dir)
            try:
                xl = _pycel(copia)
            except Exception as e:                           # noqa: BLE001
                fallos.append(fname + ': pycel no compila ('
                              + type(e).__name__ + ')')
                continue
            pref = "'" + ws.title + "'!"
            total = _ev(xl, pref + coord_total) if coord_total else None
            avance = _ev(xl, pref + coord_avance) if coord_avance else None
            suma_sub = None
            if subtot:
                vals = [_ev(xl, pref + c) for c in subtot]
                if all(isinstance(v, (int, float)) for v in vals):
                    suma_sub = round(sum(vals), 2)
            # el avance tiene que MOVERSE al marcar un ítem como terminal
            avance2 = None
            if coord_avance:
                estado_col = None
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(row=4, column=c).value
                           or ws.cell(row=3, column=c).value
                           or '').strip().lower() == 'estado':
                        estado_col = c
                if estado_col:
                    from openpyxl.utils import get_column_letter as gcl
                    primera = gcl(estado_col) + '5'
                    terminal = motor.opciones_dv(ws, primera)
                    if terminal:
                        _ev(xl, pref + coord_avance)
                        _ev(xl, pref + primera)
                        xl.set_value(pref + primera, terminal[-1])
                        avance2 = _ev(xl, pref + coord_avance)
            fila = {'fichero': fname, 'hoja': ws.title,
                    'total': total, 'avance': avance,
                    'avance_tras_marcar_1': avance2,
                    'subtotales': len(subtot),
                    'suma_subtotales': suma_sub,
                    'cuadra_subtotales': (
                        None if (suma_sub is None
                                 or not isinstance(total, (int, float)))
                        else abs(suma_sub - total) < 0.01)}
            fuera.append(fila)
            if coord_total and not isinstance(total, (int, float)):
                fallos.append(fname + ':' + ws.title + '!' + coord_total
                              + ': el TOTAL no evalúa (' + repr(total) + ')')
            if coord_avance and not isinstance(avance, (int, float)):
                fallos.append(fname + ':' + ws.title + '!' + coord_avance
                              + ': el % completado no evalúa ('
                              + repr(avance) + ')')
            if (isinstance(avance, (int, float))
                    and isinstance(avance2, (int, float))
                    and avance2 <= avance):
                fallos.append(fname + ':' + ws.title
                              + ': el % completado NO sube al marcar un ítem '
                                'como terminal (' + str(avance) + ' → '
                              + str(avance2) + '): ¿estado terminal mal leído?')
            if fila['cuadra_subtotales'] is False:
                fallos.append(fname + ':' + ws.title + ': los subtotales por '
                              'categoría suman ' + str(suma_sub)
                              + ' y el TOTAL ' + str(total))
    return {'checklists': fuera, 'fallos': fallos}


def demo_isnumber(demos_dir):
    """§1.6/§7-bis.13 — sin `ISNUMBER`, Excel ordena el texto por encima de
    cualquier número: `"No alcanzado"<0` es VERDADERO y el semáforo pinta de
    rojo justo la celda que dice que no hay dato. Con la guarda, no."""
    os.makedirs(demos_dir, exist_ok=True)
    path = os.path.join(demos_dir, '_isnumber.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'F'
    ws['A1'] = -50
    ws['A2'] = 'No alcanzado'
    ws['A3'] = ''
    ws['B1'] = '=AND(ISNUMBER(A1),A1<0)'
    ws['B2'] = '=AND(ISNUMBER(A2),A2<0)'
    ws['C2'] = '=A2<0'
    ws['B3'] = '=AND(ISNUMBER(A3),A3<0)'
    wb.save(path)
    xl = _pycel(path)
    return {'numero_negativo_con_guarda': _ev(xl, 'F!B1'),
            'texto_con_guarda': _ev(xl, 'F!B2'),
            'texto_SIN_guarda': _ev(xl, 'F!C2'),
            'vacia_con_guarda': _ev(xl, 'F!B3'),
            'ok': (_ev(xl, 'F!B1') is True and _ev(xl, 'F!B2') is False)}


def demo_sin_dato(demos_dir):
    """§7-bis.13 — «sin dato» es `""`, nunca `0`: un mes sin una sola venta no
    tiene un margen del «0,0 %»."""
    os.makedirs(demos_dir, exist_ok=True)
    path = os.path.join(demos_dir, '_sindato.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'F'
    ws['A1'] = 0
    ws['A2'] = 1000
    ws['B1'] = motor.iferror('A2/A1')
    ws['B2'] = '=IF(A1=0,0,A2/A1)'
    wb.save(path)
    xl = _pycel(path)
    return {'iferror_devuelve': repr(_ev(xl, 'F!B1')),
            'el_patron_viejo_devuelve': repr(_ev(xl, 'F!B2')),
            'ok': _ev(xl, 'F!B1') == ''}


def demo_proteccion(carpeta, nombres):
    """§1.8 — hojas protegidas SIN contraseña y con los verdes desbloqueados.

    Una celda verde que quede BLOQUEADA es una celda que el cliente no puede
    rellenar: con la protección activada, el verde deja de ser cosmético.
    """
    fuera, fallos = [], []
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            verdes = libres = 0
            for row in ws.iter_rows():
                for c in row:
                    if c.__class__.__name__ == 'MergedCell':
                        continue
                    if motor.es_verde(c) and c.data_type != 'f':
                        verdes += 1
                        if not c.protection.locked:
                            libres += 1
            clave = getattr(ws.protection, 'password', None)
            fuera.append({'fichero': fname, 'hoja': ws.title,
                          'protegida': bool(ws.protection.sheet),
                          'con_contrasena': bool(clave),
                          'verdes': verdes, 'verdes_desbloqueadas': libres})
            if not ws.protection.sheet:
                fallos.append(fname + ':' + ws.title + ': sin proteger')
            if clave:
                fallos.append(fname + ':' + ws.title + ': CON contraseña')
            if verdes != libres:
                fallos.append(fname + ':' + ws.title + ': '
                              + str(verdes - libres)
                              + ' celdas verdes BLOQUEADAS')
    return {'hojas': fuera, 'fallos': fallos, 'total_hojas': len(fuera),
            'protegidas': sum(1 for h in fuera if h['protegida'])}


def demo_bio_version(carpeta, nombres, pid):
    """§1.10/§1.11 — bio anclada (INSERCIÓN: no la llevaba ninguno de los 111)
    y línea de versión 2.0, una sola vez por fichero, más el `subject`."""
    fuera, fallos = [], []
    esperado = motor.version_line(pid)
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        bio = version = malas = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if not isinstance(v, str):
                        continue
                    if v == motor.BIO_LINE:
                        bio += 1
                    if motor.RX_VERSION.match(v):
                        version += 1
                        if v != esperado:
                            malas += 1
        subject = wb.properties.subject
        fuera.append({'fichero': fname, 'bio': bio, 'version': version,
                      'version_no_2_0': malas, 'subject': subject})
        if bio != 1:
            fallos.append(fname + ': bio ' + str(bio) + ' veces (debe ser 1)')
        if version != 1:
            fallos.append(fname + ': línea de versión ' + str(version)
                          + ' veces (debe ser 1)')
        if malas:
            fallos.append(fname + ': ' + str(malas)
                          + ' líneas de versión que no son la 2.0 de ' + pid)
        if not (isinstance(subject, str) and subject.endswith('· v2.0')):
            fallos.append(fname + ': subject = ' + repr(subject))
    return {'ficheros': fuera, 'fallos': fallos}


def demo_a4(carpeta, nombres):
    """§1.13 — la Fase A dejó el A4 puesto en las 141 hojas y el motor NO lo
    toca. Esto no escribe: comprueba que sigue ahí (incluida la hoja
    `Instrucciones` que el motor crea, que nace con su propio A4)."""
    fallos, hojas = [], 0
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        c = motor.contadores(wb, fname)
        hojas += c['hojas']
        for h in c['a4_mal']:
            fallos.append(fname + ':' + h + ': A4 incompleto')
    return {'hojas': hojas, 'fallos': fallos}


# ==========================================================================
# main
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 de la familia Guías Cómo Montar')
    ap.add_argument('--producto', default='guia-restaurante-gastronomico',
                    help='pid de la guía (carpeta de astro-site/public/dl/)')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    ap.add_argument('--solo', default='a,b,c',
                    help='grupos a aplicar: motor (ninguno), a, b, c o a,b')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--origen', default=None,
                    help='carpeta de origen alternativa (exige --dry-run)')
    ap.add_argument('--sin-idempotencia', action='store_true')
    ap.add_argument('--sin-demos', action='store_true')
    args = ap.parse_args()

    pid = args.producto
    if pid not in PRODUCTOS:
        raise SystemExit(
            'ABORTADO: «' + pid + '» no es una de las 8 guías de la familia. '
            'Son: ' + ', '.join(PRODUCTOS) + '. (Un pid mal tecleado haría '
            'que la línea de versión enlazase a una landing que no existe.)')
    if args.origen and not args.dry_run:
        raise SystemExit('ABORTADO: --origen sólo se admite con --dry-run.')
    origen = args.origen or os.path.join(DL, pid)
    destino = os.path.join(SCRATCH, 'dryrun', pid)
    idem_dir = os.path.join(SCRATCH, 'dryrun', pid + '-idem')
    demos_dir = os.path.join(SCRATCH, 'dryrun', '_demos', pid)

    if not args.dry_run and os.environ.get('GUIAS_APPLY') != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en ' + origen
            + ', que la SPEC declara intocable hasta que T5 firme. Usa '
              '--dry-run (o GUIAS_APPLY=1 si eres el orquestador y ya tienes '
              'el visto bueno).')

    motor.CTX['producto'] = pid
    respaldo = None
    if args.dry_run:
        preparar_copia(origen, destino)
        carpeta = destino
    else:
        carpeta = origen
        respaldo = os.path.join(
            SCRATCH, 'respaldos', pid + '.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(origen, respaldo)
        log('  respaldo previo de los entregables: ' + respaldo)

    # ---- grupos y contenido ---------------------------------------------
    grupos, cargados, ausentes, avisos = [], [], [], []
    pedidos = [g.strip().lower() for g in args.solo.split(',') if g.strip()]
    solo_motor = all(l in ('motor', 'ninguno', 'none', '') for l in pedidos)
    contenidos = {}
    for letra in pedidos:
        if letra in ('motor', 'ninguno', 'none', ''):
            continue
        try:
            g = importlib.import_module('grupo_' + letra)
        except ImportError as e:                             # noqa: BLE001
            # Un grupo PEDIDO que no carga es ROJO: un paquete a medio
            # construir no puede declararse terminado.
            ausentes.append('grupo_' + letra + ' se pidió y NO se pudo cargar '
                            '(' + str(e) + '): las secciones que le tocan '
                            'quedan SIN aplicar')
            log('  grupo_' + letra + ' no disponible: ' + str(e))
            continue
        grupos.append((g, letra))
        cargados.append(letra)
        contenidos[letra] = cargar_contenido(pid, letra, avisos)
        log('  grupo_' + letra + ' cargado'
            + (' + contenido' if contenidos.get(letra) else ' (sin contenido)'))

    nombres = ficheros_de(carpeta)
    if not nombres:
        raise SystemExit('0 xlsx en ' + carpeta)

    # ---- 1/7 pre-vuelo de moldes ----------------------------------------
    log('\n== 1/7 · pre-vuelo de moldes (§1.1) sobre ' + str(len(nombres))
        + ' xlsx ==')
    detectado, fallos_molde = prevuelo(carpeta, nombres, pid)
    for d in detectado:
        for ch in d['checklists']:
            log('  ' + d['fichero'] + ':' + ch['hoja'] + ' → molde '
                + ch['molde'] + ' (cabecera fila '
                + str(ch['fila_cabecera']) + ')')
        if d['variante_pl']:
            log('  ' + d['fichero'] + ' → variante_pl ' + d['variante_pl'])
    if fallos_molde:
        log('\nABORTADO — molde no reconocido:')
        for x in fallos_molde:
            log('  ' + x)
        informe = {'producto': pid, 'version': '2.0', 'rol': 'motor',
                   'spec': SPEC, 'abortado_por': 'molde_desconocido',
                   'modo': 'dry-run' if args.dry_run else 'produccion',
                   'fallos': fallos_molde, 'exit': 2}
        if args.json:
            os.makedirs(os.path.dirname(os.path.abspath(args.json)),
                        exist_ok=True)
            with open(args.json, 'w', encoding='utf-8') as fh:
                json.dump(informe, fh, ensure_ascii=False, indent=1)
        return 2

    nombres = ficheros_a_tocar(nombres, grupos, pid)
    informe_ficheros, registros = [], {}

    # ---- 2/7 post-proceso ------------------------------------------------
    log('\n== 2/7 · post-proceso de ' + str(len(nombres)) + ' ficheros ==')
    for fname in nombres:
        registros[fname] = procesar(carpeta, fname, grupos, contenidos,
                                    informe_ficheros, pid)
        log('  ' + fname + ': ' + str(len(registros[fname]))
            + ' fórmulas nuevas')

    # ---- 3/7 idempotencia ------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 3/7 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(idem_dir):
            shutil.rmtree(idem_dir)
        shutil.copytree(carpeta, idem_dir)
        antes = dict((n, digest(os.path.join(carpeta, n))) for n in nombres)
        for fname in nombres:
            procesar(idem_dir, fname, grupos, contenidos, None, pid)
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(idem_dir, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:60]}
        log('  diferencias 1.ª vs 2.ª pasada: ' + str(len(difs)))
        for d in difs[:12]:
            log('    ' + d)

    # ---- 4/7 inject_cache (al final del todo) ----------------------------
    log('\n== 4/7 · inject_cache ==')
    cache = inject_cache(carpeta, nombres)

    # ---- 5/7 verificación data_only --------------------------------------
    log('\n== 5/7 · verificación data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log('  con valor: ' + str(ver['con_valor']) + ' · "" por diseño: '
        + str(ver['vacias_por_diseno']) + ' · fallos: '
        + str(len(ver['fallos'])))
    for fl in ver['fallos'][:12]:
        log('    ' + fl)

    # ---- 6/7 censo -------------------------------------------------------
    log('\n== 6/7 · censo-entregables --fail ==')
    cen = censo(carpeta)

    # ---- 7/7 demostraciones ---------------------------------------------
    demos = {}
    if not args.sin_demos:
        log('\n== 7/7 · demostraciones (pycel) ==')
        demos = {
            'checklists': demo_checklists(carpeta, nombres, demos_dir),
            'semaforo_isnumber': demo_isnumber(demos_dir),
            'sin_dato_es_vacio': demo_sin_dato(demos_dir),
            'proteccion': demo_proteccion(carpeta, nombres),
            'bio_y_version': demo_bio_version(carpeta, nombres, pid),
            'a4_intacto': demo_a4(carpeta, nombres),
        }
        log('  checklists con bloque de resumen: '
            + str(len(demos['checklists']['checklists'])))
        for ch in demos['checklists']['checklists']:
            if ch.get('molde_sin_bloque'):
                log('    ' + ch['fichero'] + ': molde B sin bloque; contador '
                    'propio intacto (' + str(ch['contador_propio_intacto'])
                    + ' fórmulas COUNTIF)')
                continue
            log('    ' + ch['fichero'] + ': total=' + str(ch.get('total'))
                + ' avance=' + str(ch.get('avance')) + ' → '
                + str(ch.get('avance_tras_marcar_1'))
                + ' subtot=' + str(ch.get('subtotales'))
                + ' cuadra=' + str(ch.get('cuadra_subtotales')))
        log('  hojas protegidas: ' + str(demos['proteccion']['protegidas'])
            + '/' + str(demos['proteccion']['total_hojas']))

    # ---- veredicto -------------------------------------------------------
    fallos = list(ausentes)
    for g, letra in grupos:
        if hasattr(g, 'demos'):
            propias = _llamar(g.demos, carpeta=carpeta, origen=origen,
                              destino=demos_dir,
                              contenido=contenidos.get(letra)) or {}
            fallos += propias.pop('fallos', [])
            demos.update(propias)
    if idem.get('diferencias'):
        fallos.append('idempotencia: ' + str(idem['diferencias'])
                      + ' diferencias entre la 1.ª y la 2.ª pasada')
    if ver['fallos']:
        fallos.append('cache: ' + str(len(ver['fallos']))
                      + ' fórmulas nuevas sin valor')
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió ' + str(cen['exit'])
                      + ': ' + ' | '.join(cen['salida'][-4:]))
    for nombre, salida, rc in cache:
        if rc != 0:
            fallos.append('inject_cache falló en ' + nombre + ' (exit '
                          + str(rc) + ')')
    if demos:
        fallos += demos.get('checklists', {}).get('fallos', [])
        fallos += demos.get('proteccion', {}).get('fallos', [])
        fallos += demos.get('bio_y_version', {}).get('fallos', [])
        fallos += demos.get('a4_intacto', {}).get('fallos', [])
        if not demos.get('semaforo_isnumber', {}).get('ok'):
            fallos.append('§1.6: la guarda ISNUMBER no discrimina texto de '
                          'número — el semáforo pintaría el aviso de rojo')
        if not demos.get('sin_dato_es_vacio', {}).get('ok'):
            fallos.append('§7-bis.13: `iferror()` no devuelve "" sino '
                          + str(demos['sin_dato_es_vacio']))

    literales_total = sum(fi.get('literales_sospechosos_total', 0)
                          for fi in informe_ficheros)
    muestra = []
    for fi in informe_ficheros:
        for lit in fi.get('literales_sospechosos', [])[:3]:
            muestra.append(fi['fichero'] + ':' + lit['hoja'] + '!'
                           + lit['celda'] + '=' + lit['literal'])
    # AVISO, no fallo: los literales vivos son deuda de §2/§4 (el `0.28` del
    # food cost y el `1.10` del IVA), que corrigen los GRUPOS. Con `--solo
    # motor` tienen que seguir ahí; convertirlos en fallo haría rojo el motor
    # por un defecto que no le toca.
    avisos_literales = ('§1.3: ' + str(literales_total) + ' literales dentro '
                        'de fórmulas que deberían ser parámetro en celda ('
                        + ', '.join(muestra[:10]) + ')') if literales_total \
        else None
    if avisos_literales:
        if solo_motor:
            avisos.append(avisos_literales)
        else:
            fallos.append(avisos_literales)

    informe = {
        'producto': pid,
        'version': '2.0',
        'rol': 'motor',
        'spec': SPEC,
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'carpeta_origen': origen,
        'carpeta_trabajo': carpeta,
        'grupos_pedidos': pedidos,
        'grupos_cargados': cargados,
        'grupos_ausentes': ausentes,
        'contenido_cargado': dict((k, bool(v)) for k, v in
                                  contenidos.items()),
        'solo_motor': solo_motor,
        'moldes_detectados': detectado,
        'literales_sospechosos_total': literales_total,
        'ficheros': informe_ficheros,
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'data_only_formulas_nuevas': ver,
            'censo_entregables': cen,
        },
        'demostraciones': demos,
        'avisos': avisos,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1)
        log('\ninforme → ' + args.json)

    for a in avisos:
        log('  AVISO ' + a)
    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'TODO VERDE (moldes, idempotencia, cache, censo, demos)'))
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (BÓRRALO sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
