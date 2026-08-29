#!/usr/bin/env python3
"""
motor.py — Motor común de la familia «Guías Cómo Montar» v2.0.

Implementa el **§1** de `guias-v2-SPEC.md` sobre los **111 xlsx** de las **8**
guías de `astro-site/public/dl/guia-*`. NO toca ficheros: recibe un `Workbook`
ya cargado y lo modifica en memoria; quien guarda es `main.py`.

El motor **no sabe de negocio**: detecta el molde, aplica las convenciones de
familia y ofrece las primitivas que usan `grupo_a/b/c`. Todo lo que dependa del
concepto de la guía (nombres de hoja, filas, cifras) vive en
`contenido_<pid>/a.py|b.py|c.py`.

Qué cubre (§1.1-§1.13):
  1.1  `molde_checklist()` y `variante_pl()` — detección ANTES de escribir, con
       **aborto** (`MoldeDesconocido`) si no reconoce. Nunca «molde A por
       defecto» (§7-bis.11).
  1.2  `a_formula()` — constante → fórmula conservando el número como dato de
       ejemplo en la celda de entrada.
  1.3  `retirar_verde_de_calculadas()` + `verdes_por_dv()` + el mapa de
       columnas de entrada por molde: verde ⇔ desbloqueada ⇔ el cliente escribe.
  1.4  `formato_por_etiqueta()` — formato por TIPO de dato leído de la etiqueta
       (fila o cabecera de columna), nunca por bloque.
  1.5  `nota_iva()` + `PARAMETROS['iva_restauracion']` — el IVA declarado y en
       celda, nunca `*1.10` incrustado.
  1.6  `semaforo_isnumber()` / `semaforo_texto()` — toda guarda de formato
       condicional que pueda leer texto lleva `ISNUMBER`.
  1.7  `dv_lista/dv_numerica/dv_porcentaje/dv_fecha`, todas con
       `showErrorMessage=True`.
  1.8  `proteger()` — protección de hoja SIN contraseña, verdes desbloqueadas.
  1.9  `cerrar_checklist()` — fila TOTAL, subtotales por categoría y
       `% completado`, respetando el TOTAL que los moldes C y D **ya traen**.
  1.10 `hoja_instrucciones()` (la crea donde falta: 56 de los 111 no la tienen),
       renombra la pestaña `Sheet`, y `cierre_instrucciones()` escribe la línea
       de versión 2.0 + la **bio anclada** (medido: 0 de 111 la llevan).
  1.11 `metadatos()` — `title`/`subject` → `… · v2.0`.
  1.12 `ensanchar_etiquetas()` + `arreglar_cabecera_escandallo()`.
  1.13 Lo que el motor NO hace: no crea ni renombra FICHEROS, no toca
       `paperSize`/márgenes/pie de las hojas que ya existen (Fase A), no escribe
       `externalLink`, no borra filas de datos.

CONVENCIONES (`cerrar()` las garantiza): editables verdes `E8F5E9` y
desbloqueadas, calculadas sin relleno; parámetros en celda, nunca literales;
`IFERROR` en toda división; «sin dato» = `""`, nunca `0`; DV con
`showErrorMessage=True`; protección sin contraseña; bio anclada;
«Versión 2.0 · agosto 2026»; metadata `… · v2.0`; ninguna celda con `''` (el
censo la cuenta como defecto `empty_str`).

pycel 1.0b30 (medido, SPEC cabecera): evalúa `SUM`, `SUMPRODUCT`, `SUMIF`,
`COUNTIF`, `IFERROR`, `IF`/`AND`, `TEXT`, `NPV`, `ROUND`, `MATCH`+`INDEX`.
**NO** implementa `IRR`, `PMT` ni `COUNTA`: `COUNTA(r)` → `COUNTIF(r,"<>")`; la
cuota de un préstamo va como anualidad algebraica `importe*i/(1-(1+i)^-n)`.

IDEMPOTENCIA: todo lo que escribe el motor es **absoluto** (mismo valor en la
misma celda, calculado a partir de la última fila de datos MEDIDA, no de dónde
estaba el bloque en la pasada anterior). Formato condicional y validaciones se
PURGAN antes de reescribirse — es donde openpyxl acumula duplicados en la 2.ª
pasada. El bloque de cierre de los checklists y el de `Instrucciones` se borran
enteros y se reescriben al final del contenido vivo.

⚠️ Los `.md`/`.xlsx` de la familia pueden llevar ESPACIO FINO (U+202F) y GUION
NO SEPARABLE (U+2011). Aquí se referencian SIEMPRE por escape (`NARROW`,
`NOBRK`), nunca escribiendo el carácter: al pasar por un heredoc del shell
degeneran en espacio y guion normales y ninguna sustitución encuentra su patrón.
"""
import copy
import re

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# Caracteres tipográficos de la familia, SIEMPRE por escape (ver docstring).
NARROW = '\u202f'          # espacio fino antes de las unidades
NOBRK = '\u2011'           # guion no separable en los rangos

# ==========================================================================
# Contexto del producto en curso (lo fija main.py antes de procesar)
# ==========================================================================
CTX = {'producto': None}

# ==========================================================================
# Paleta, formatos y textos fijos
# ==========================================================================
VERDE = 'E8F5E9'          # celda editable (verde de edición de la familia)
CF_VERDE_BG, CF_VERDE_FG = 'C6EFCE', '006100'
CF_AMBAR_BG, CF_AMBAR_FG = 'FFEB9C', '9C6500'
CF_ROJO_BG, CF_ROJO_FG = 'FFC7CE', '9C0006'
CF_GRIS_BG, CF_GRIS_FG = 'F2F2F2', '595959'

FMT_EUR = '#,##0.00 €'
FMT_PCT = '0.0%'
FMT_ENT = '#,##0'
FMT_FECHA = 'dd/mm/yyyy'

BIO_LINE = ('Diseñado por John Guerrero — chef y consultor gastronómico desde '
            '2010, en cocina desde los 17 años · johnguerrero.es')
RX_BIO = re.compile(r'John\s+Guerrero')

RX_VERSION = re.compile(r'^Versi[óo]n\s+\d+\.\d+\s+·')

NOTA_DESPROTEGER = ('Para editar la estructura o una celda que no esté en '
                    'verde: Revisar → Desproteger hoja (no tiene contraseña).')
RX_DESPROTEGER = re.compile(r'^Para editar (la estructura|una celda)')

NOTA_VERDES = 'Celdas verdes = campos editables'

# §1.5(b) — la base de las cifras, dicha en voz alta en cada libro.
NOTA_IVA = ('Todas las cifras van SIN IVA. En el cash flow también se '
            'escriben sin IVA: su capa de IVA (modelo 303) lo añade aparte, '
            'porque la tesorería es caja.')
#: En la propia hoja de tesorería la frase general se lee al revés («todas van
#: sin IVA, salvo el cash flow» escrito DENTRO del cash flow), así que ahí va la
#: variante que dice lo que aplica a esa hoja.
#: RT-06/RD-15 · la primera redacción ordenaba teclear los importes CON IVA
#: («Las cifras de esta hoja van CON IVA: es caja») mientras la nota del bloque
#: 303, tres filas más abajo, decía lo contrario. Quien hiciera caso de la
#: cabecera tecleaba importes con IVA y la capa se lo volvía a sumar: el IVA
#: repercutido contado dos veces y el break-even adelantado. Una sola frase, y
#: la que describe lo que el libro hace de verdad.
NOTA_IVA_CAJA = ('Escribe los ingresos y los gastos SIN IVA, igual que en el '
                 'P&L: la capa de IVA (modelo 303) de más abajo lo añade, '
                 'porque el cash flow es caja.')
RX_NOTA_IVA = re.compile(r'^(Todas las cifras|Las cifras de esta hoja)')
RX_HOJA_CAJA = re.compile(r'cash\s*flow|flujo de caja|tesorer', re.I)

# Bloque de cierre que el motor DUEÑA en los checklists (§1.9). El marcador
# delimita la región que se borra y se reescribe entera en cada pasada: así el
# bloque nunca se duplica ni se desplaza.
MARCA_BLOQUE = 'RESUMEN — lo calcula el libro (v2.0)'
#: Etiqueta de la fila que cuenta las partidas SIN importe (RT-09/RD-33) y
#: valor de la columna que decide si una fila entra en el TOTAL (RD-18/RC-23).
ETIQUETA_SIN_TASAR = 'Partidas SIN importe (a presupuestar)'
COL_INCLUIR = 'Incluir en el total'
INCLUIR_SI, INCLUIR_NO = 'Sí', 'No'
RX_MARCA = re.compile(r'^RESUMEN — lo calcula el libro')

RX_PIE_MARCA = re.compile(r'^AI Chef Pro\s+·\s+aichef\.pro')

ETIQUETA_TOTAL = 'TOTAL PRESUPUESTADO (€)'
ETIQUETA_AVANCE = 'Avance del checklist (% completado)'
ETIQUETA_SUBTOT = 'SUBTOTALES POR CATEGORÍA'

# §1.5 / §7-bis.3 / §7-bis.16 — parámetros normativos SIEMPRE en celda verde y
# con nota. Ninguna cifra se teclea dentro de una fórmula.
PARAMETROS = {
    'iva_restauracion': {
        'etiqueta': 'Tipo de IVA de restauración (%)',
        'valor': 0.10,
        'formato': FMT_PCT,
        'nota': ('10 % general en restauración; 21 % en bebidas alcohólicas. '
                 'Cámbialo aquí y se recalcula todo el libro.'),
    },
    'ss_empresa': {
        'etiqueta': 'Seguridad Social a cargo de la empresa (%)',
        'valor': 0.33,
        'formato': FMT_PCT,
        'nota': ('Cotización empresarial aproximada sobre el salario bruto '
                 '(contingencias comunes, desempleo, FOGASA y formación). '
                 'Ajústala a tu convenio y a tus contratos.'),
    },
    'smi_anual': {
        'etiqueta': 'SMI vigente (€/año, 14 pagas)',
        'valor': 17094.0,
        'formato': FMT_EUR,
        'nota': ('1.221 €/mes × 14 pagas = 17.094 €/año — Real Decreto '
                 '126/2026, de 18 de febrero (BOE-A-2026-3815), con efectos '
                 'desde el 1-ene-2026. El convenio provincial de hostelería '
                 'prevalece si fija un mínimo superior.'),
    },
}
SMI_ANUAL = PARAMETROS['smi_anual']['valor']
IVA_RESTAURACION = PARAMETROS['iva_restauracion']['valor']
SS_EMPRESA = PARAMETROS['ss_empresa']['valor']


def version_line(pid=None):
    """`Versión 2.0 · agosto 2026 · aichef.pro/<pid> · info@aichef.pro` (§1.10).

    El `pid` sale del contexto que fija `main.py`: la familia son 8 productos y
    cada uno enlaza a SU landing.
    """
    pid = pid or CTX.get('producto')
    if not pid:
        raise RuntimeError('motor.CTX["producto"] sin fijar: la línea de '
                           'versión enlazaría a una landing equivocada')
    return ('Versión 2.0 · agosto 2026 · aichef.pro/' + pid
            + ' · info@aichef.pro')


class MoldeDesconocido(Exception):
    """§1.1/§7-bis.11 — el motor NO adivina un molde. Aborta con el fichero."""


# ==========================================================================
# Registro de fórmulas (main.py verifica una por una que quedaron cacheadas)
# ==========================================================================
REGISTRO = []


def reg(ws, coord, formula):
    REGISTRO.append((ws.title, coord, formula))


def f(ws, coord, formula, fmt=None, align=None, bold=None):
    """Escribe una FÓRMULA y la registra para la verificación `data_only`."""
    cel = ws[coord]
    cel.value = formula
    if fmt:
        cel.number_format = fmt
    if align:
        cel.alignment = Alignment(horizontal=align)
    if bold is not None:
        cel.font = Font(bold=bold, size=cel.font.size, color=cel.font.color)
    reg(ws, coord, formula)
    return cel


def val(ws, coord, valor, fmt=None, verde_=False, bold=None, align=None,
        wrap=None):
    """Escribe un VALOR (constante). `verde_` lo marca como editable."""
    cel = ws[coord]
    cel.value = valor
    if fmt:
        cel.number_format = fmt
    if verde_:
        cel.fill = PatternFill('solid', fgColor=VERDE)
        cel.protection = Protection(locked=False)
    if bold is not None:
        cel.font = Font(bold=bold, size=cel.font.size, color=cel.font.color)
    if align or wrap:
        cel.alignment = Alignment(horizontal=align or 'general',
                                  vertical='top', wrap_text=bool(wrap))
    return cel


def iferror(expresion, alterna=''):
    """`=IFERROR(<expr>,"<alterna>")` — guarda obligatoria en TODA división.

    `alterna` vacía produce `""`, que es lo que la familia entiende por «sin
    dato» (§7-bis.13): nunca un `0`, que en un margen se lee como «0,0 %».
    """
    expr = expresion[1:] if expresion.startswith('=') else expresion
    return '=IFERROR(' + expr + ',"' + alterna + '")'


# ==========================================================================
# Verdes, bloqueo y limpieza (§1.3)
# ==========================================================================
def es_verde(cel):
    relleno = cel.fill
    return (relleno is not None and relleno.fill_type == 'solid'
            and relleno.fgColor is not None
            and isinstance(relleno.fgColor.rgb, str)
            and relleno.fgColor.rgb.upper().endswith(VERDE))


def _celdas(ws, rango):
    return ws[rango] if ':' in rango else [[ws[rango]]]


def _rango_combinado(ws, coord):
    """Coordenadas del rango COMBINADO al que pertenece `coord`, o `()`."""
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            return tuple(c.coordinate
                         for fila in ws[str(mr)] for c in fila)
    return ()


def verde(ws, rango):
    """Marca en verde (y DESBLOQUEA) un rango 'A1:C3' o una celda 'A1'.

    RT-26: si la celda encabeza un rango COMBINADO, el relleno se aplica a
    todas sus celdas. `escandallo-maestro!'Ficha (plantilla)'!C4:D4` se veía
    medio verde y medio blanco, y D4 quedaba desbloqueada sin verde — la única
    excepción del producto a la regla §1.3 «verde ⇔ desbloqueada».
    """
    for fila in _celdas(ws, rango):
        for cel in fila:
            if cel.__class__.__name__ == 'MergedCell':
                continue
            cel.fill = PatternFill('solid', fgColor=VERDE)
            cel.protection = Protection(locked=False)
            for coord in _rango_combinado(ws, cel.coordinate):
                otra = ws[coord]
                otra.fill = PatternFill('solid', fgColor=VERDE)
                otra.protection = Protection(locked=False)


def quitar_verde(ws, rango):
    """Retira el verde de una celda que ha pasado a ser CALCULADA (§1.3).

    Hoy hay 45 celdas verdes en `calculadora-ticket-medio!B5:D20` y otras 45 en
    `pl-mensual-escenarios!'Escenarios'` incluidas las **filas de resultado**
    (TEC-01, TEC-02): la hoja le pide al cliente que teclee él el resultado.
    """
    for fila in _celdas(ws, rango):
        for cel in fila:
            if cel.__class__.__name__ == 'MergedCell':
                continue
            cel.fill = PatternFill()
            cel.protection = Protection(locked=True)


def marcar_editable(ws, rango):
    """Desbloquea por ROL, no por color: hay inputs pintados con el color de su
    escenario. `proteger()` respeta esta marca aunque el relleno diga otra cosa.
    """
    marcadas = getattr(ws, '_g_editables', None)
    if marcadas is None:
        marcadas = set()
        ws._g_editables = marcadas
    for fila in _celdas(ws, rango):
        for cel in fila:
            marcadas.add(cel.coordinate)
            if cel.__class__.__name__ != 'MergedCell':
                cel.protection = Protection(locked=False)


def permitir_negativo(ws, rango):
    """Excluye un rango del «≥ 0» de `validaciones()`: un EBITDA presupuestado,
    un flujo acumulado o un saldo en descubierto TIENEN que poder ser
    negativos, y una DV bloqueante ahí empuja al cliente a desproteger la hoja
    el primer día — justo lo que la protección venía a evitar."""
    libres = getattr(ws, '_g_negativos', None)
    if libres is None:
        libres = set()
        ws._g_negativos = libres
    for fila in _celdas(ws, rango):
        for cel in fila:
            libres.add(cel.coordinate)


def limpiar_rango(ws, rango):
    """Vacía un rango (valor, relleno, formato, bloqueo) deshaciendo antes las
    combinaciones que lo tocan: una `MergedCell` tiene el `value` de sólo
    lectura y asignarle nada revienta con AttributeError."""
    objetivo = CellRange(rango if ':' in rango else rango + ':' + rango)
    for m in [str(r) for r in ws.merged_cells.ranges]:
        try:
            CellRange(m).intersection(objetivo)
        except ValueError:
            continue
        ws.unmerge_cells(m)
    for fila in _celdas(ws, rango):
        for cel in fila:
            if cel.__class__.__name__ == 'MergedCell':
                continue
            cel.value = None
            cel.fill = PatternFill()
            cel.number_format = 'General'
            cel.protection = Protection(locked=True)
            cel.font = Font()
            cel.alignment = Alignment()


def retirar_verde_de_calculadas(ws, informe=None, fname=''):
    """§1.3 — regla dura de la familia: si la celda la calcula el libro, no
    puede estar pintada de verde ni desbloqueada. Se ejecuta en `cerrar()`,
    DESPUÉS de los grupos, para que alcance a lo que ellos hayan convertido en
    fórmula."""
    n = 0
    editables = getattr(ws, '_g_editables', set())
    for row in ws.iter_rows():
        for c in row:
            if c.data_type != 'f':
                continue
            if c.coordinate in editables:
                continue
            if es_verde(c):
                c.fill = PatternFill()
                n += 1
            c.protection = Protection(locked=True)
    if n and informe is not None:
        informe.append(fname + ':' + ws.title + ': ' + str(n)
                       + ' celdas calculadas dejan de estar en verde (§1.3)')
    return n


def verdes_por_dv(ws, informe=None, fname=''):
    """Toda celda con desplegable es un input: verde y desbloqueada (§1.3+§1.8).

    Sin esto la protección del §1.8 deja al cliente con un desplegable que no
    puede desplegar — y ése es el caso REAL de `checklist-apertura-legal!G`
    (dark-kitchen) y de las 26 casillas `✓/☐/N/A` de cada checklist de
    panadería, que hoy no llevan verde ninguno.
    """
    n = 0
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for rango in dv.sqref.ranges:
            for fila in ws[str(rango)] if ':' in str(rango) \
                    else [[ws[str(rango)]]]:
                for cel in fila:
                    if cel.__class__.__name__ == 'MergedCell':
                        continue
                    if cel.data_type == 'f':
                        continue
                    if not es_verde(cel):
                        cel.fill = PatternFill('solid', fgColor=VERDE)
                        n += 1
                    cel.protection = Protection(locked=False)
    if n and informe is not None:
        informe.append(fname + ':' + ws.title + ': ' + str(n)
                       + ' celdas con desplegable pasan a verde/editable')
    return n


# ==========================================================================
# §1.2 — constantes → fórmulas conservando el número
# ==========================================================================
def a_formula(ws, coord, formula, celda_ejemplo=None, fmt=None, informe=None,
              fname='', nota=None):
    """Convierte una celda que hoy trae una CONSTANTE en una fórmula.

    §1.2/§7-bis.12: el número que había **no se pierde**. Si se indica
    `celda_ejemplo` (la celda de entrada de la que la fórmula depende), el valor
    viejo se deposita allí como dato de ejemplo; si no, se anota en el informe
    para que el orquestador lo vea. Nunca se borra un número que el cliente
    pueda estar usando sin sustituirlo por algo que calcule lo mismo.

    Devuelve el valor anterior (o `None` si la celda ya era fórmula o estaba
    vacía) para que quien llama pueda comprobar la tolerancia de 0,01 € con
    pycel.
    """
    cel = ws[coord]
    anterior = cel.value if cel.data_type != 'f' else None
    if anterior is not None and celda_ejemplo is not None:
        destino = ws[celda_ejemplo]
        if destino.value is None:
            destino.value = anterior
            destino.fill = PatternFill('solid', fgColor=VERDE)
            destino.protection = Protection(locked=False)
    f(ws, coord, formula, fmt=fmt)
    if informe is not None and anterior is not None:
        informe.append(
            fname + ':' + ws.title + '!' + coord + ': constante '
            + repr(anterior) + ' → fórmula ' + formula[:70]
            + (' (ejemplo conservado en ' + celda_ejemplo + ')'
               if celda_ejemplo else ' (VALOR ANTERIOR SIN DESTINO: '
               'compruébalo contra pycel)')
            + (' · ' + nota if nota else ''))
    return anterior


def escribir_parametro(ws, fila, col_etiqueta, col_valor, clave,
                       col_nota=None, valor=None):
    """§1.5/§7-bis.3/§7-bis.16 — un parámetro normativo (IVA, SS, SMI) SIEMPRE
    en celda verde y con su nota al lado. Nunca `*1.10` dentro de la fórmula.
    """
    p = PARAMETROS[clave]
    val(ws, col_etiqueta + str(fila), p['etiqueta'])
    val(ws, col_valor + str(fila), p['valor'] if valor is None else valor,
        fmt=p['formato'], verde_=True)
    if col_nota:
        val(ws, col_nota + str(fila), p['nota'], wrap=True)
    return col_valor + str(fila)


# ==========================================================================
# §1.6 — formato condicional, SIEMPRE purgando antes de escribir
# ==========================================================================
def _norm_rango(ref):
    try:
        return CellRange(ref).coord
    except Exception:                                        # noqa: BLE001
        return ref


def _limpiar_cf(ws, rango):
    """Quita las reglas cuyo sqref es exactamente `rango`, comparado
    NORMALIZADO: openpyxl guarda `C5:C5` como `C5`, así que comparar cadenas
    sueltas no encuentra la regla anterior y la 2.ª pasada apila una copia (el
    fichero se ve igual en Excel y la idempotencia salta)."""
    objetivo = _norm_rango(rango)
    supervivientes = []
    for cf in ws.conditional_formatting:
        if _norm_rango(str(cf.sqref)) == objetivo:
            continue
        supervivientes.append((str(cf.sqref), list(cf.rules)))
    nueva = ConditionalFormattingList()
    for sqref, reglas in supervivientes:
        for r in reglas:
            nueva.add(sqref, r)
    ws.conditional_formatting = nueva


def _dxf(bg, fg):
    return DifferentialStyle(font=Font(color=fg, bold=True),
                             fill=PatternFill(start_color=bg, end_color=bg,
                                              fill_type='solid'))


def semaforo_texto(ws, rango, vocabulario):
    """Colorea por TEXTO EXACTO (§3.3). `vocabulario` = ((texto, bg, fg), …).

    RT-27/RC-33: la primera versión usaba `containsText` + `SEARCH`, que es
    subcadena e insensible a mayúsculas, mientras el contador de avance de la
    misma columna usa `COUNTIF(rango,"Completado")`, que exige la celda
    completa. Un «Completado parcialmente» se pintaba de verde y no contaba
    como completado. `cellIs equal` es la comparación que casa con la DV de
    lista cerrada que ya gobierna esas columnas.
    """
    _limpiar_cf(ws, rango)
    for texto, bg, fg in vocabulario:
        regla = Rule(type='cellIs', operator='equal',
                     formula=['"' + texto + '"'],
                     dxf=_dxf(bg, fg), stopIfTrue=True)
        ws.conditional_formatting.add(rango, regla)
    return True


def semaforo_isnumber(ws, rango, ancla, operador='<', umbral='0',
                      bg=CF_ROJO_BG, fg=CF_ROJO_FG, parar=True):
    """§1.6 — ninguna regla compara directamente una celda que puede traer `""`
    o un texto de aviso: la guarda es `=AND(ISNUMBER($X5),$X5<0)`.

    Sin `ISNUMBER`, Excel evalúa `""<0` como FALSO y `"No alcanzado"<0` como
    VERDADERO (el texto ordena por encima de cualquier número): el semáforo
    pintaría de rojo justo la celda que dice que no hay dato.
    """
    formula = ('=AND(ISNUMBER(' + ancla + '),' + ancla + operador + umbral
               + ')')
    _limpiar_cf(ws, rango)
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=parar,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))
    return formula


def regla_expresion(ws, rango, formula, bg=CF_ROJO_BG, fg=CF_ROJO_FG,
                    parar=True):
    _limpiar_cf(ws, rango)
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=parar,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


# ==========================================================================
# §1.7 — validación de datos, también purgando
# ==========================================================================
def _coords_de_dv(ws, dv):
    fuera = set()
    for rango in dv.sqref.ranges:
        ref = str(rango)
        for fila in ws[ref] if ':' in ref else [[ws[ref]]]:
            for cel in fila:
                fuera.add(cel.coordinate)
    return fuera


def _purgar_dv(ws, tipo, coords):
    """Elimina las DV de `tipo` que solapen con `coords`."""
    fuera = []
    for dv in ws.data_validations.dataValidation:
        if dv.type == tipo and (_coords_de_dv(ws, dv) & set(coords)):
            continue
        fuera.append(dv)
    ws.data_validations.dataValidation = fuera


def _purgar_dv_duplicadas(ws):
    vistas, fuera = set(), []
    for dv in ws.data_validations.dataValidation:
        clave = (dv.type, dv.formula1, dv.formula2, dv.operator,
                 str(dv.sqref))
        if clave in vistas:
            continue
        vistas.add(clave)
        fuera.append(dv)
    ws.data_validations.dataValidation = fuera


def _ordenar(coords):
    return sorted(set(coords),
                  key=lambda c: (int(re.search(r'(\d+)', c).group(1)),
                                 column_index_from_string(
                                     re.match(r'([A-Z]+)', c).group(1))))


def dv_lista(ws, coords, opciones, titulo='Valor no válido', mensaje=None):
    """Desplegable con `showErrorMessage=True` (§1.7)."""
    coords = _ordenar(coords)
    if not coords:
        return None
    _purgar_dv(ws, 'list', coords)
    dv = DataValidation(type='list',
                        formula1='"' + ','.join(opciones) + '"',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle=titulo,
                        error=mensaje or ('Elige un valor de la lista: '
                                          + ', '.join(opciones)))
    ws.add_data_validation(dv)
    for c in coords:
        dv.add(c)
    _purgar_dv_duplicadas(ws)
    return dv


def dv_numerica(ws, coords, minimo=0, maximo=None, titulo=None, mensaje=None,
                prompt=None):
    coords = _ordenar(coords)
    if not coords:
        return None
    _purgar_dv(ws, 'decimal', coords)
    if maximo is None:
        dv = DataValidation(type='decimal', operator='greaterThanOrEqual',
                            formula1=str(minimo), allow_blank=True,
                            showErrorMessage=True,
                            errorTitle=titulo or 'Importe no válido',
                            error=mensaje or ('Escribe un número mayor o igual '
                                              'que ' + str(minimo) + '.'))
    else:
        dv = DataValidation(type='decimal', operator='between',
                            formula1=str(minimo), formula2=str(maximo),
                            allow_blank=True, showErrorMessage=True,
                            errorTitle=titulo or 'Valor fuera de rango',
                            error=mensaje or ('Escribe un valor entre '
                                              + str(minimo) + ' y '
                                              + str(maximo) + '.'))
    if prompt:
        dv.showInputMessage = True
        dv.promptTitle = (titulo or 'Cómo se escribe')[:32]
        dv.prompt = prompt
    ws.add_data_validation(dv)
    for c in coords:
        dv.add(c)
    _purgar_dv_duplicadas(ws)
    return dv


#: RT-15/RC-11 · texto por DEFECTO de la DV de porcentaje. El de la merma
#: dejó de ser el genérico: explicaba una división por `1-merma` al hacer clic
#: en el tipo de IVA, en la SS de la empresa o en el interés del préstamo.
PCT_DEFECTO = ('Porcentaje', 'Se escribe en tanto por uno: 0,20 = 20 %.', 1.0)
#: RT-16 · el tope 0,95 SÓLO tiene sentido donde el porcentaje entra en una
#: división `1-x`. En un «% de comensales» impedía teclear el 100 % que el
#: propio semáforo de la hoja exige.
PCT_MERMA = ('Merma (%)',
             'Se escribe en tanto por uno: 0,20 = 20 %. Tope 0,95 porque la '
             'merma entra en el coste como neta/(1-merma).', 0.95)


def prompt_porcentaje(ws, rango, titulo, prompt, maximo=1.0):
    """Registra el texto (y el tope) de la DV de porcentaje de un rango.

    Se marca por ROL, igual que `_g_editables` o `_g_formatos`: `validaciones()`
    agrupa las celdas verdes de porcentaje por el texto registrado y crea una DV
    por grupo, en vez de pegar la misma frase en las 10 hojas del producto.
    """
    reg = getattr(ws, '_g_pct', None)
    if reg is None:
        reg = {}
        ws._g_pct = reg
    for fila in _celdas(ws, rango):
        for cel in fila:
            reg[cel.coordinate] = (titulo, prompt, maximo)
    return rango


def dv_porcentaje(ws, coords, titulo=None, prompt=None, maximo=None):
    """§1.7 — porcentajes: decimal 0-`maximo` con mensaje de entrada propio."""
    t = titulo or PCT_DEFECTO[0]
    p = prompt or PCT_DEFECTO[1]
    m = PCT_DEFECTO[2] if maximo is None else maximo
    return dv_numerica(
        ws, coords, minimo=0, maximo=m, titulo=t,
        mensaje='Escribe un porcentaje entre 0 y ' + ('%g' % m).replace('.', ',')
                + ' (0,20 = 20 %).',
        prompt=p)


def dv_fecha(ws, coords):
    coords = _ordenar(coords)
    if not coords:
        return None
    _purgar_dv(ws, 'date', coords)
    dv = DataValidation(type='date', operator='between',
                        formula1='DATE(2020,1,1)', formula2='DATE(2040,12,31)',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle='Fecha no válida',
                        error='Escribe una fecha (dd/mm/aaaa) entre 2020 y '
                              '2040.')
    ws.add_data_validation(dv)
    for c in coords:
        dv.add(c)
    _purgar_dv_duplicadas(ws)
    return dv


def dv_propia(ws, rango, minimo, maximo, titulo, mensaje, prompt=None):
    """Registra una DV numérica PROPIA para un rango (RT-14).

    `validaciones()` clasifica las celdas verdes por formato y sólo distingue
    «importe», «porcentaje» y «admite negativo». Las dos entradas que gobiernan
    la proyección a 3 años caían en el último cajón y acababan con un
    `decimal >= -1.000.000.000.000`, que no valida nada: teclear «8» pensando
    en 8 % multiplicaba los ingresos del año 2 por nueve sin un solo aviso.
    """
    reg = getattr(ws, '_g_dv_prop', None)
    if reg is None:
        reg = {}
        ws._g_dv_prop = reg
    for fila in _celdas(ws, rango):
        for cel in fila:
            reg[cel.coordinate] = (minimo, maximo, titulo, mensaje, prompt)
    return rango


def validaciones(ws, informe=None, fname=''):
    """DV numérica sobre las celdas verdes, clasificadas por FORMATO de número.

    `General` (texto libre: un nombre, una nota) se queda sin validación, y las
    que ya tienen desplegable o fecha tampoco se tocan.
    """
    con_lista = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type in ('list', 'date'):
            con_lista |= _coords_de_dv(ws, dv)
    libres = set(getattr(ws, '_g_negativos', set()))
    propias = getattr(ws, '_g_dv_prop', {})
    grupos_prop = {}
    importes, porcentajes, negativos = [], [], []
    for row in ws.iter_rows():
        for c in row:
            if not es_verde(c) or c.coordinate in con_lista:
                continue
            if c.data_type == 'f':
                continue
            fmt = c.number_format or ''
            if 'yy' in fmt or 'YY' in fmt:
                continue
            if c.coordinate in propias:
                grupos_prop.setdefault(propias[c.coordinate], []).append(
                    c.coordinate)
            elif c.coordinate in libres:
                negativos.append(c.coordinate)
            elif '%' in fmt:
                porcentajes.append(c.coordinate)
            elif '€' in fmt or fmt.startswith('#,##0'):
                importes.append(c.coordinate)
    for (mini, maxi, titulo, mensaje, prompt), coords in grupos_prop.items():
        dv_numerica(ws, coords, minimo=mini, maximo=maxi, titulo=titulo,
                    mensaje=mensaje, prompt=prompt)
    if importes:
        dv_numerica(ws, importes, minimo=0)
    if porcentajes:
        # RT-15/RC-11 + RT-16: una DV por TEXTO registrado, no una sola para
        # todo el producto. Sin registro, el genérico (0-1).
        reg = getattr(ws, '_g_pct', {})
        grupos = {}
        for c in porcentajes:
            grupos.setdefault(reg.get(c, PCT_DEFECTO), []).append(c)
        for (titulo, prompt, maximo), coords in grupos.items():
            dv_porcentaje(ws, coords, titulo=titulo, prompt=prompt,
                          maximo=maximo)
    if negativos:
        dv_numerica(ws, negativos, minimo=-1000000000000,
                    titulo='Importe',
                    mensaje='Escribe un número (puede ser negativo).')
    if informe is not None and (importes or porcentajes or negativos
                                or grupos_prop):
        informe.append(fname + ':' + ws.title + ': DV en '
                       + str(len(importes)) + ' importes, '
                       + str(len(porcentajes)) + ' porcentajes, '
                       + str(len(negativos)) + ' que admiten negativo y '
                       + str(sum(len(v) for v in grupos_prop.values()))
                       + ' con rango propio')
    return len(importes), len(porcentajes)


# ==========================================================================
# §1.4 — formato por TIPO de dato, leído de la etiqueta
# ==========================================================================
RX_PCT = re.compile(r'\(\s*%\s*\)|%\s*$|^%\s|\bmargen\b.*\(%\)', re.I)
RX_FECHA_ET = re.compile(r'^(fecha|inicio|fin)\b|fecha l[íi]mite', re.I)
RX_RECUENTO = re.compile(
    r'\b(uds|unidades|uds\.?\s*vendidas|cubiertos|comensales|d[íi]as|'
    r'pedidos|tickets|plazas|personas|piezas|raciones|meses|n[ºo]\b)', re.I)
RX_IMPORTE = re.compile(r'\(\s*€\s*\)|\b€\b|\(eur\)', re.I)
#: RT-19/RC-24 · etiquetas que son DINERO aunque no traigan «(€)». La regla de
#: §1.4 decide por el texto, y dos filas de IVA con el rótulo largo quedaban en
#: `#,##0` (sin símbolo y sin decimales) junto a trece vecinas en `#,##0.00 €`.
RX_DINERO = re.compile(
    r'\b(iva|cuota|cuotas|liquidaci[oó]n|importe|importes|coste|costes|'
    r'gasto|gastos|ingreso|ingresos|cobro|cobros|pago|pagos|amortizaci[oó]n|'
    r'n[oó]mina|n[oó]minas|salario|bruto|presupuesto|inversi[oó]n|fianza|'
    r'renta|alquiler|facturaci[oó]n|ebitda|ebit|tesorer[ií]a|desembolso)\b',
    re.I)


def fijar_formato(ws, coord, fmt):
    """§1.4 — CLAVA el formato de una celda concreta frente a la regla de
    columna. Mismo canal que `_g_editables`/`_g_negativos`: el grupo marca por
    ROL y `cerrar()` lo respeta.

    Nace de un caso medido: `plan-financiero-3-anos!'P&L Mensual'!B35`
    («Margen EBITDA») es un RATIO dentro de una columna cuya cabecera dice
    `Importe (€)`. La pasada (a) de `formato_por_etiqueta` aplica el tipo de la
    cabecera a toda la columna, así que sin este pin el motor devolvería B35 a
    `#,##0.00 €` justo después de que el grupo lo pusiera en `0.0%`, y el
    cliente seguiría leyendo «0,19 €» donde pone 18,6 % (TEC-09). Igual pasa
    con la celda de meses del fondo de maniobra en `Inversión`, cuya columna C
    se rotula `Presupuesto (€)`.
    """
    fijados = getattr(ws, '_g_formatos', None)
    if fijados is None:
        fijados = {}
        ws._g_formatos = fijados
    fijados[coord] = fmt
    ws[coord].number_format = fmt
    return coord


def tipo_por_etiqueta(texto):
    """Devuelve el formato que la etiqueta IMPONE, o `None` si no lo dice.

    Se mira en orden: `%` gana a `€` (una «Margen EBITDA (%)» hoy va en
    `#,##0.00 €`), fecha gana a recuento, y el recuento gana al importe (los
    «Cubiertos/día» y las «Uds Vendidas» van hoy en euros: TEC-24, TEC-25).
    """
    if not isinstance(texto, str):
        return None
    t = texto.strip()
    if not t:
        return None
    if RX_PCT.search(t):
        return FMT_PCT
    if RX_FECHA_ET.search(t):
        return FMT_FECHA
    if RX_RECUENTO.search(t) and not RX_IMPORTE.search(t):
        return FMT_ENT
    if RX_IMPORTE.search(t):
        return FMT_EUR
    if RX_DINERO.search(t):
        return FMT_EUR
    return None


def formato_ficha(ws, informe=None, fname='', ultima=5):
    """§1.4 en hojas SIN cabecera de tabla: la etiqueta va en la columna A y el
    valor a su derecha (fichas tipo `Break-Even` o el `Ticket Medio` de los
    hermanos, donde `A5='Ticket medio (€)'` y `B5=45` sin ninguna fila de
    encabezado). Sin esto, `'Break-Even'!B6` («Comensales/día») y `B7` («Días
    abierto/mes») se quedarían en `#,##0.00 €` (TEC-25).

    Sólo toca filas cuya etiqueta DECLARA el tipo y celdas que no contienen
    texto: una nota escrita al lado no se convierte en número.
    """
    cambios, detalle = 0, []
    for r in range(1, ws.max_row + 1):
        fmt = tipo_por_etiqueta(ws.cell(row=r, column=1).value)
        if not fmt:
            continue
        for c in range(2, min(ultima, ws.max_column) + 1):
            cel = ws.cell(row=r, column=c)
            if cel.__class__.__name__ == 'MergedCell':
                continue
            if cel.coordinate in getattr(ws, '_g_formatos', {}):
                continue
            if isinstance(cel.value, str):
                continue
            if cel.number_format != fmt:
                detalle.append(cel.coordinate + ':' + cel.number_format + '→'
                               + fmt)
                cel.number_format = fmt
                cambios += 1
    if cambios and informe is not None:
        informe.append(fname + ':' + ws.title + ': ' + str(cambios)
                       + ' celdas reformateadas en modo ficha (§1.4) '
                       + str(detalle[:8]))
    return cambios


def formato_por_etiqueta(ws, fila_cab, informe=None, fname=''):
    """§1.4 — formato por tipo, no por bloque, en dos pasadas complementarias.

    **Por columna**: si la cabecera de la columna dice el tipo («Coste Est.
    (€)», «Uds Vendidas», «Fecha Límite»), manda sobre toda la columna.
    **Por fila**: si la etiqueta de la columna A dice el tipo («Food cost (%)»,
    «Días abierto/mes») y la cabecera de las columnas de valor NO dice ninguno
    («Pesimista», «Escenario 1», «Mes 3»), manda la fila.

    Las dos nunca compiten: la de fila se inhibe en las columnas cuya cabecera
    ya lleva marca de tipo. Sin esa regla, en `plan-financiero!'P&L Mensual'`
    (cabeceras «Importe (€)» y «% s/Ventas») una etiqueta de fila reformatearía
    la columna de porcentajes.
    """
    if not fila_cab:
        return 0
    ncol = ws.max_column
    cab = {}
    for c in range(1, ncol + 1):
        cab[c] = ws.cell(row=fila_cab, column=c).value
    tipo_col = dict((c, tipo_por_etiqueta(v)) for c, v in cab.items())
    cambios = 0
    detalle = []

    # (a) por columna
    for c in range(2, ncol + 1):
        fmt = tipo_col.get(c)
        if not fmt:
            continue
        for r in range(fila_cab + 1, ws.max_row + 1):
            cel = ws.cell(row=r, column=c)
            if cel.__class__.__name__ == 'MergedCell':
                continue
            if cel.coordinate in getattr(ws, '_g_formatos', {}):
                continue
            if cel.number_format != fmt:
                detalle.append(cel.coordinate + ':' + cel.number_format
                               + '→' + fmt + ' [col ' + str(cab.get(c)) + ']')
                cel.number_format = fmt
                cambios += 1

    # (b) por fila, sólo donde la cabecera de columna no manda
    columnas_libres = [c for c in range(2, ncol + 1)
                       if not tipo_col.get(c)
                       and not (isinstance(cab.get(c), str)
                                and cab[c].strip().lower()
                                in ('notas', 'nota', 'observaciones'))]
    if columnas_libres:
        for r in range(fila_cab + 1, ws.max_row + 1):
            etiqueta = ws.cell(row=r, column=1).value
            fmt = tipo_por_etiqueta(etiqueta)
            if not fmt:
                continue
            for c in columnas_libres:
                cel = ws.cell(row=r, column=c)
                if cel.__class__.__name__ == 'MergedCell':
                    continue
                if cel.coordinate in getattr(ws, '_g_formatos', {}):
                    continue
                if cel.number_format != fmt:
                    detalle.append(cel.coordinate + ':' + cel.number_format
                                   + '→' + fmt + ' [fila '
                                   + str(etiqueta)[:28] + ']')
                    cel.number_format = fmt
                    cambios += 1
    if cambios and informe is not None:
        informe.append(fname + ':' + ws.title + ': ' + str(cambios)
                       + ' celdas reformateadas por tipo de dato desde la '
                       'cabecera de la fila ' + str(fila_cab) + ' (§1.4) '
                       + str(detalle[:8]))
    return cambios


# ==========================================================================
# §1.12 — anchos y cabeceras cortadas
# ==========================================================================
def anchos(ws, mapa):
    for col, w in mapa.items():
        if ws.column_dimensions[col].width != w:
            ws.column_dimensions[col].width = w


def ensanchar_etiquetas(ws, informe=None, fname='', margen=2, tope=60):
    """§1.12/TEC-22 — ninguna etiqueta más larga que su columna con la contigua
    ocupada. Sólo ENSANCHA (nunca estrecha): así es idempotente y no deshace un
    ancho puesto a mano.
    """
    cambios = []
    ncol = ws.max_column
    for row in ws.iter_rows():
        for cel in row:
            v = cel.value
            if not isinstance(v, str) or len(v) < 8:
                continue
            if cel.data_type == 'f':
                # El `.value` de una celda de fórmula ES la fórmula. Sin este
                # corte, `=IF('P&L Mensual'!B10=0,"",…)` cuenta como una
                # etiqueta de 50 caracteres y ensancha su columna a 50: medido
                # en la hoja «Proyección 3 Años» del representante y en 35
                # columnas del cash flow japonés.
                continue
            if cel.__class__.__name__ == 'MergedCell':
                continue
            if any(cel.coordinate in CellRange(str(m))
                   for m in ws.merged_cells.ranges):
                continue          # una etiqueta combinada ya desborda a gusto
            col = cel.column
            if col >= ncol:
                continue
            vecina = ws.cell(row=cel.row, column=col + 1)
            if vecina.value is None:
                continue          # puede desbordar sin tapar nada
            L = get_column_letter(col)
            actual = ws.column_dimensions[L].width or 8.43
            necesario = min(tope, len(v) + margen)
            if necesario > actual:
                ws.column_dimensions[L].width = necesario
                cambios.append(L + ':' + str(round(actual, 1)) + '→'
                               + str(necesario) + ' (' + v[:30] + ')')
    if cambios and informe is not None:
        informe.append(fname + ':' + ws.title + ': ' + str(len(cambios))
                       + ' columnas ensanchadas (§1.12) ' + str(cambios[:6]))
    return cambios


def arreglar_cabecera_escandallo(ws, informe=None, fname=''):
    """§1.12/TEC-22, caso medido del representante: `escandallo-maestro!A4`
    («Nombre del plato:», 17 caracteres) vive en la columna A con `width=5.0` y
    con `B4:D4` combinada a su derecha, así que no puede desbordar. Se combina
    `A4:B4` para la etiqueta y `C4:D4` como campo del cliente.

    Guardado por FIRMA exacta: si el fichero no es ése (los hermanos tienen otra
    rejilla), no se toca nada.
    """
    if ws['A4'].value != 'Nombre del plato:':
        return False
    merges = set(str(m) for m in ws.merged_cells.ranges)
    if 'A4:B4' in merges and 'C4:D4' in merges:
        return False                                  # ya arreglado
    if 'B4:D4' in merges:
        ws.unmerge_cells('B4:D4')
    ws.merge_cells('A4:B4')
    ws.merge_cells('C4:D4')
    verde(ws, 'C4')
    if ws.column_dimensions['G'].width and ws.column_dimensions['G'].width < 22:
        ws.column_dimensions['G'].width = 22
    if informe is not None:
        informe.append(fname + ':' + ws.title
                       + ': A4:B4 etiqueta + C4:D4 campo; G a 22 (§1.12)')
    return True


# ==========================================================================
# §1.1 — DETECCIÓN DE MOLDE (antes de escribir nada; aborta si no reconoce)
# ==========================================================================
def cabeceras(ws, fila):
    fuera = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila, column=c).value
        fuera[get_column_letter(c)] = v.strip() if isinstance(v, str) else v
    return fuera


def _norm(v):
    return v.strip() if isinstance(v, str) else v


#: Firmas MEDIDAS el 2026-08-29 abriendo los 111 xlsx con openpyxl. La regla
#: literal del §1.1 de la SPEC reconoce 40 de los 46 checklists de la familia:
#: falla en los 5 checklists de panadería que no son `checklist-legal` (su D3
#: no es «Plazo orientativo» sino «Registro», «Responsable», «Precio min €»,
#: «Coste orientativo» o «Plazo») y en `guia-dark-kitchen/
#: checklist-apertura-legal.xlsx`, que NO es el molde C: su F4 es «Coste
#: Estimado (€)» y su G4 «Estado» — el molde A con coste y estado
#: INTERCAMBIADOS. El censo de T0 ya los dejó marcados como
#: `checklist_atipico_no_reconocido_por_regla_1.1` y sin fila de cabecera.
#: Aplicarles el molde A escribiría el TOTAL sobre la columna «Estado».
#: Por eso aquí hay CUATRO moldes y la firma de B es estructural (la casilla),
#: no una etiqueta concreta.


def molde_checklist(ws, fname='', pid=''):
    """Devuelve 'A' | 'B' | 'C' | 'D' mirando la cabecera (fila 3 o 4).

    · **A** (38 ficheros MEDIDOS: 8 del representante + 6 en cada uno de los 5
      hermanos —casual, mexicano, peruano, japonés, nikkei—): fila 4,
      `B='Categoría'`, `F='Estado'`, `G='Coste Est. (€)'`. La SPEC §3.3 dice
      «14 ficheros: 8 del representante + 6 por hermano ×5», que son 38, no 14.
    · **B** (6 de panadería): fila 3, `A='✓'` y casillas `☐` en la columna A.
    · **C** (`guia-dark-kitchen/checklist-equipamiento-obra.xlsx`): fila 4,
      `E='Presupuesto (€)'`, `G='Desviación (%)'`.
    · **D** (`guia-dark-kitchen/checklist-apertura-legal.xlsx`): fila 4,
      `F='Coste Estimado (€)'`, `G='Estado'`.

    Si no encaja en ninguno **aborta con el nombre del fichero y la cabecera
    cruda**: nunca «aplica el molde A por defecto» (§1.1, §7-bis.11).
    """
    for fila in (4, 3):
        c = cabeceras(ws, fila)
        if c.get('B') == 'Categoría' and c.get('G') == 'Coste Est. (€)':
            return 'A', fila
        if c.get('A') == '✓':
            casillas = sum(1 for r in range(fila + 1, ws.max_row + 1)
                           if _norm(ws.cell(row=r, column=1).value)
                           in ('☐', '✓', 'N/A'))
            if casillas:
                return 'B', fila
        if c.get('E') == 'Presupuesto (€)' and c.get('G') == 'Desviación (%)':
            return 'C', fila
        if c.get('F') == 'Coste Estimado (€)' and c.get('G') == 'Estado':
            return 'D', fila
    crudo = dict((k, v) for k, v in cabeceras(ws, 4).items() if v is not None)
    crudo3 = dict((k, v) for k, v in cabeceras(ws, 3).items() if v is not None)
    raise MoldeDesconocido(
        (pid + '/' if pid else '') + fname + ':' + ws.title
        + ': cabecera no reconocida por §1.1. fila4=' + repr(crudo)
        + ' fila3=' + repr(crudo3)
        + ' — el motor NO aplica un molde por defecto: añade la firma a '
          'motor.molde_checklist() con la evidencia medida.')


def es_checklist(fname):
    return 'checklist' in fname.lower()


def fila_cabecera_tabla(ws, tope=6, minimo=3):
    """Fila que encabeza la TABLA de datos, o `None` si no hay una clara.

    ⚠️ Este detector nació de un fallo propio, medido: la primera versión se
    conformaba con «la primera fila (4 o 3) con 3 o más textos», y en
    `escandallo-maestro.xlsx` la fila 4 es la banda «Nombre del plato: · E4
    Raciones: · G4 Food Cost Objetivo: · H4 28%» — no la cabecera, que está en
    la 6. Con esa fila, `formato_por_etiqueta` leía «Raciones:» como recuento y
    ponía `#,##0` a la columna `Precio/Ud (€)`, y leía «28%» y ponía `0.0%` a
    `Coste (€)`. El regex no falló: acertó en el sitio equivocado.

    Criterio, verificable de un vistazo: gana la fila con MÁS textos
    CONSECUTIVOS empezando en la columna A (una cabecera de tabla no tiene
    huecos por la izquierda), con al menos `minimo`, y a igualdad la más alta.
    Si ninguna llega al mínimo devuelve `None` y el motor no reformatea nada:
    prefiere no tocar a tocar por aproximación.
    """
    mejor, mejor_n = None, 0
    for r in range(1, min(tope, ws.max_row) + 1):
        n = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                n += 1
            else:
                break
        if n > mejor_n:
            mejor, mejor_n = r, n
    return mejor if mejor_n >= minimo else None


def variante_pl(wb):
    """§1.1 — tres variantes MEDIDAS de `pl-mensual-escenarios.xlsx`:

    · `'escenarios-columnas'` — una hoja `Escenarios` con los tres escenarios
      en las columnas B/C/D (representante).
    · `'tres-hojas'` — `Pesimista`/`Realista`/`Optimista` (los 5 hermanos,
      casual incluido), con constantes tecleadas y 0 fórmulas.
    · `'hoja-unica'` — `P&L 3 escenarios` (panadería).
    """
    hojas = set(wb.sheetnames)
    if {'Pesimista', 'Realista', 'Optimista'} <= hojas:
        return 'tres-hojas'
    if 'Escenarios' in hojas:
        return 'escenarios-columnas'
    if 'P&L 3 escenarios' in hojas:
        return 'hoja-unica'
    return None


def opciones_dv(ws, coord):
    """Opciones del desplegable que cubre `coord`, en su orden original.

    Se lee del fichero y NO se da por supuesta: el estado terminal es
    «Completado» en los moldes A y D, pero **«Instalado»** en
    `checklist-equipamiento-obra` (dark-kitchen), cuya lista es
    `Pendiente,Pedido,Recibido,Instalado`. Un `% completado` que buscara
    «Completado» ahí devolvería 0 % para siempre.
    """
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        if coord in _coords_de_dv(ws, dv):
            crudo = (dv.formula1 or '').strip()
            if crudo.startswith('"') and crudo.endswith('"'):
                return [x.strip() for x in crudo[1:-1].split(',')]
    return []


# ==========================================================================
# §1.9/§1.10 — cierre de los checklists
# ==========================================================================
#: Columnas de ENTRADA por molde: lo que el cliente rellena. Con la protección
#: del §1.8 activada, una columna de entrada que no quede verde/desbloqueada es
#: una columna que el cliente NO puede rellenar — el verde deja de ser
#: cosmético y pasa a ser el mapa de lo editable.
ENTRADAS = {
    'A': ('D', 'E', 'F', 'G', 'H'),   # Responsable, Fecha, Estado, Coste, Notas
    'C': ('D', 'E', 'F', 'H', 'I'),   # Proveedor, Presup., Real, Estado, Notas
    'D': ('D', 'E', 'F', 'G', 'H'),   # Responsable, Fecha, Coste, Estado, Notas
}
#: Molde B: sólo la casilla (que ya lleva su DV) y la columna de notas. No se
#: le añade columna de coste (§7-bis.17): sería inventar ~200 importes.

SEM_ESTADO = ((('Completado'), CF_VERDE_BG, CF_VERDE_FG),
              (('Instalado'), CF_VERDE_BG, CF_VERDE_FG),
              (('Recibido'), CF_AMBAR_BG, CF_AMBAR_FG),
              (('En Curso'), CF_AMBAR_BG, CF_AMBAR_FG),
              (('En Proceso'), CF_AMBAR_BG, CF_AMBAR_FG),
              (('Pedido'), CF_AMBAR_BG, CF_AMBAR_FG),
              (('Pendiente'), CF_GRIS_BG, CF_GRIS_FG))


def _fin_datos(ws, molde, fila_cab):
    """Última fila de DATOS, MEDIDA en tiempo de ejecución (§9, gate de
    recuento): los recuentos del R1 no cuadran entre sí — su inventario dice
    «checklist-legal 42 filas» y lo medido son 40 (A5:A44)."""
    fin = fila_cab
    for r in range(fila_cab + 1, ws.max_row + 1):
        if molde == 'B':
            if _norm(ws.cell(row=r, column=1).value) in ('☐', '✓', 'N/A'):
                fin = r
            elif ws.cell(row=r, column=2).value is not None:
                fin = r          # fila de sección (encabezado combinado)
        else:
            if isinstance(ws.cell(row=r, column=1).value, (int, float)) \
                    and not isinstance(ws.cell(row=r, column=1).value, bool):
                fin = r
    return fin


def _fila_marca(ws):
    """Fila donde empieza el bloque que el motor DUEÑA, si ya está escrito."""
    for r in range(1, ws.max_row + 1):
        v = _norm(ws.cell(row=r, column=1).value)
        if isinstance(v, str) and RX_MARCA.match(v):
            return r
    return None


def _fila_total_existente(ws, fin, marca=None):
    """Molde C y D **ya traen** la fila TOTAL: no se duplica (§3.3).

    ⚠️ La búsqueda se corta en `MARCA_BLOQUE`. Sin ese corte, en la 2.ª pasada
    el motor encontraba su PROPIO `TOTAL PRESUPUESTADO (€)`, lo tomaba por el
    total nativo del libro y empezaba el bloque dos filas más abajo: la
    idempotencia daba 168 diferencias y el bloque bajaba dos filas en cada
    ejecución.
    """
    tope = marca if marca else ws.max_row + 1
    for r in range(fin + 1, min(fin + 6, tope - 1) + 1):
        for c in range(1, ws.max_column + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if isinstance(v, str) and v.upper().startswith('TOTAL'):
                return r
    return None


def _col(letra):
    return column_index_from_string(letra)


def cerrar_checklist(ws, molde, fila_cab, fname='', informe=None,
                     pid=None):
    """§1.9 + §1.10 + §3.3 sobre los tres (cuatro) moldes.

    Escribe SIEMPRE un bloque completo al final, delimitado por `MARCA_BLOQUE`:
    en cada pasada se borra desde el inicio calculado hasta la última fila y se
    reescribe idéntico, así que la 2.ª pasada no lo duplica ni lo desplaza. El
    inicio se calcula desde la última fila de datos MEDIDA, nunca desde donde
    estaba el bloque antes.
    """
    cab = cabeceras(ws, fila_cab)
    fin = _fin_datos(ws, molde, fila_cab)
    if fin <= fila_cab:
        if informe is not None:
            informe.append(fname + ':' + ws.title
                           + ': sin filas de datos, no se cierra')
        return None
    ncol = ws.max_column
    ultima_col = get_column_letter(max(ncol, 5))

    # --- columnas por rol, leídas de la cabecera (no por posición fija) ----
    col_estado = col_coste = col_cat = col_fecha = None
    for L, txt in cab.items():
        if not isinstance(txt, str):
            continue
        t = txt.strip().lower()
        if t == 'estado':
            col_estado = L
        elif t in ('coste est. (€)', 'coste estimado (€)', 'presupuesto (€)'):
            col_coste = L
        elif t in ('categoría', 'categoria', 'zona'):
            col_cat = L
        elif t in ('fecha límite', 'fecha limite'):
            col_fecha = L

    # --- entradas verdes + DV + formato (§1.3, §1.4, §1.7) ----------------
    if molde in ENTRADAS:
        for L in ENTRADAS[molde]:
            if _col(L) > ncol:
                continue
            rango = L + str(fila_cab + 1) + ':' + L + str(fin)
            hay_formula = any(ws[L + str(r)].data_type == 'f'
                              for r in range(fila_cab + 1, fin + 1))
            if hay_formula:
                continue          # columna calculada (Desviación del molde C)
            verde(ws, rango)
    if col_fecha:
        for r in range(fila_cab + 1, fin + 1):
            ws[col_fecha + str(r)].number_format = FMT_FECHA
        dv_fecha(ws, [col_fecha + str(r)
                      for r in range(fila_cab + 1, fin + 1)])
    if col_coste:
        for r in range(fila_cab + 1, fin + 1):
            cel = ws[col_coste + str(r)]
            if cel.data_type != 'f':
                cel.number_format = FMT_EUR

    # --- desplegable de Estado y semáforo por texto exacto (§1.6, §3.3) ---
    terminal = None
    if col_estado:
        ancla = col_estado + str(fila_cab + 1)
        opciones = opciones_dv(ws, ancla)
        rango_estado = ancla + ':' + col_estado + str(fin)
        if opciones:
            terminal = opciones[-1]
            dv_lista(ws, [col_estado + str(r)
                          for r in range(fila_cab + 1, fin + 1)], opciones)
            vocab = [t for t in SEM_ESTADO if t[0] in opciones]
            if vocab:
                semaforo_texto(ws, rango_estado, tuple(vocab))
    if molde == 'B':
        # La columna de notas la escribe el cliente: con la protección del §1.8
        # activada, dejarla bloqueada le impediría anotar en su propio
        # checklist. Se localiza por CABECERA, no por letra: en panadería es E
        # en cuatro ficheros, y en `checklist-appcc` esa misma E es «Acción
        # correctiva» (contenido del producto, que sí queda bloqueado).
        for L, txt in cab.items():
            if isinstance(txt, str) and txt.strip().lower() in ('notas',
                                                                'nota'):
                verde(ws, L + str(fila_cab + 1) + ':' + L + str(fin))
        # La casilla ya trae su DV `"✓,☐,N/A"` y su formato condicional
        # `$A4="✓"`: se respetan (§3.3). Sólo se garantiza el
        # `showErrorMessage` y que la casilla quede editable.
        coords = []
        for r in range(fila_cab + 1, fin + 1):
            if _norm(ws.cell(row=r, column=1).value) in ('☐', '✓', 'N/A'):
                coords.append('A' + str(r))
        opciones = opciones_dv(ws, coords[0]) if coords else []
        if coords and opciones:
            dv_lista(ws, coords, opciones)

    # --- bloque de resumen -------------------------------------------------
    if molde == 'B':
        # §3.3/§7-bis.17 — panadería ya trae su contador
        # (`C36='=COUNTIF(A4:A34,"✓")'` / `E36='=COUNTIF(B4:B34,"?*")'`) y NO
        # lleva columna de coste: añadirle una obligaría a inventar ~200
        # importes sin tasar. Así que aquí no se escribe bloque ninguno — sólo
        # se ha garantizado el `showErrorMessage` de la casilla y su verde.
        if informe is not None:
            informe.append(fname + ':' + ws.title + ': molde B, '
                           + str(fin - fila_cab) + ' filas medidas; contador '
                           'propio respetado y sin fila de coste (§7-bis.17)')
        return {'molde': molde, 'fila_cabecera': fila_cab, 'fin_datos': fin,
                'items': fin - fila_cab, 'fila_total': None,
                'estado_terminal': None, 'inicio_bloque': None}
    marca = _fila_marca(ws)
    fila_total = _fila_total_existente(ws, fin, marca)
    inicio = (fila_total or fin) + 2
    pie = None
    for r in range(inicio, ws.max_row + 1):
        for c in range(1, ncol + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if isinstance(v, str) and RX_PIE_MARCA.match(v):
                pie = v
    limpiar_rango(ws, 'A' + str(inicio) + ':' + ultima_col
                  + str(max(ws.max_row, inicio)))

    r = inicio
    val(ws, 'A' + str(r), MARCA_BLOQUE, bold=True)
    r += 1
    # RD-18/RC-23 — el grupo puede haber marcado alternativas EXCLUYENTES (una
    # capota o un túnel de lavado, un bloque de cocción estándar o uno de alta
    # gama). Si las dos entran en el mismo `SUM`, el TOTAL presupuesta las dos
    # opciones a la vez en cuanto el cliente tase la segunda.
    col_inc = getattr(ws, '_g_col_incluir', None)
    ini_d, fin_d = str(fila_cab + 1), str(fin)

    def _suma(col):
        if col_inc:
            return ('=SUMIF($' + col_inc + '$' + ini_d + ':$' + col_inc + '$'
                    + fin_d + ',"' + INCLUIR_SI + '",$' + col + '$' + ini_d
                    + ':$' + col + '$' + fin_d + ')')
        return '=SUM(' + col + ini_d + ':' + col + fin_d + ')'

    if col_coste and not fila_total:
        # §1.9/TEC-16/COM-31 — los 313.290 € del representante repartidos en 8
        # listas dejan de sumarse a mano.
        val(ws, 'A' + str(r), ETIQUETA_TOTAL, bold=True)
        f(ws, col_coste + str(r), _suma(col_coste), fmt=FMT_EUR, bold=True)
        r += 1
        # RT-09/RD-33 — el TOTAL es el argumento de venta de §1.9 y se
        # entregaba sistemáticamente corto: 32 de los 91 ítems de equipamiento
        # y las nueve filas nuevas del checklist legal (entre ellas la fianza,
        # que con el alquiler del propio pack son 34.000 €) van sin importe.
        # No se inventa el precio: se DICE cuántas partidas faltan por tasar.
        val(ws, 'A' + str(r), ETIQUETA_SIN_TASAR, bold=True)
        f(ws, col_coste + str(r),
          '=COUNTBLANK(' + col_coste + ini_d + ':' + col_coste + fin_d + ')',
          fmt='#,##0', bold=True)
        fijar_formato(ws, col_coste + str(r), '#,##0')
        regla_expresion(ws, col_coste + str(r),
                        '=AND(ISNUMBER(' + col_coste + str(r) + '),'
                        + col_coste + str(r) + '>0)', bg=CF_AMBAR_BG,
                        fg=CF_AMBAR_FG)
        val(ws, get_column_letter(min(ncol, _col(col_coste) + 1)) + str(r),
            'El TOTAL de arriba es un SUELO: no incluye estas partidas, que '
            'van a presupuestar.')
        r += 1
        # RD-17 — columnas de importe DERIVADAS que el grupo ha marcado para
        # que también se totalicen (la dotación de menú degustación con
        # precio, por ejemplo): sin su total, el cliente lee que necesita el
        # doble de piezas y sigue llevando al banco el importe de la carta.
        for col_x, etiqueta in (getattr(ws, '_g_totalizar', None) or []):
            val(ws, 'A' + str(r), etiqueta, bold=True)
            f(ws, col_x + str(r), _suma(col_x), fmt=FMT_EUR, bold=True)
            r += 1
    elif fila_total and informe is not None:
        informe.append(fname + ':' + ws.title + ': TOTAL ya presente en la '
                       'fila ' + str(fila_total) + ', no se duplica (§3.3)')
    if col_estado and terminal:
        val(ws, 'A' + str(r), ETIQUETA_AVANCE, bold=True)
        rango_e = (col_estado + str(fila_cab + 1) + ':' + col_estado
                   + str(fin))
        f(ws, col_estado + str(r),
          iferror('COUNTIF(' + rango_e + ',"' + terminal + '")/COUNTIF('
                  + rango_e + ',"<>")'), fmt=FMT_PCT, bold=True)
        val(ws, get_column_letter(min(ncol, _col(col_estado) + 1)) + str(r),
            'Estado terminal de este checklist: «' + terminal + '»')
        r += 1
    # RT-28: la condición era `molde == 'A'`, una lista blanca. El molde D
    # (checklist-apertura-legal de dark-kitchen) tiene columna de categoría y
    # columna de coste y se quedaba sin desglose por una etiqueta, no por una
    # ausencia. Se condiciona a lo que la operación necesita de verdad.
    if col_cat and col_coste:
        r += 1
        val(ws, 'A' + str(r), ETIQUETA_SUBTOT, bold=True)
        r += 1
        vistas = []
        for fila in range(fila_cab + 1, fin + 1):
            v = _norm(ws[col_cat + str(fila)].value)
            if isinstance(v, str) and v and v not in vistas:
                vistas.append(v)
        for cat in vistas:
            val(ws, 'A' + str(r), cat)
            # El criterio va por REFERENCIA a la etiqueta, no como literal
            # entrecomillado: una categoría con coma o comilla rompería el
            # SUMIF y el fallo no se vería hasta abrir el fichero.
            base_sub = ('$' + col_cat + '$' + ini_d + ':$' + col_cat + '$'
                        + fin_d + ',$A' + str(r))
            if col_inc:
                # el subtotal hereda la exclusión del TOTAL: si no, la suma de
                # las categorías no cuadraría con él (§1.9).
                f(ws, col_coste + str(r),
                  '=SUMIFS($' + col_coste + '$' + ini_d + ':$' + col_coste
                  + '$' + fin_d + ',' + base_sub + ',$' + col_inc + '$' + ini_d
                  + ':$' + col_inc + '$' + fin_d + ',"' + INCLUIR_SI + '")',
                  fmt=FMT_EUR)
            else:
                f(ws, col_coste + str(r),
                  '=SUMIF(' + base_sub + ',$' + col_coste + '$' + ini_d + ':$'
                  + col_coste + '$' + fin_d + ')', fmt=FMT_EUR)
            r += 1
    if pie:
        r += 1
        val(ws, 'A' + str(r), pie)
        ws.merge_cells('A' + str(r) + ':' + ultima_col + str(r))
    if informe is not None:
        informe.append(fname + ':' + ws.title + ': molde ' + molde
                       + ', cabecera fila ' + str(fila_cab) + ', '
                       + str(fin - fila_cab) + ' ítems medidos, bloque de '
                       'resumen en A' + str(inicio))
    return {'molde': molde, 'fila_cabecera': fila_cab, 'fin_datos': fin,
            'items': fin - fila_cab, 'fila_total': fila_total,
            'estado_terminal': terminal, 'inicio_bloque': inicio}


# ==========================================================================
# §1.10 — Instrucciones, pestaña, versión y bio
# ==========================================================================
RX_ILEGAL_HOJA = re.compile(r'[\\/\?\*\[\]:]')

#: Partículas con las que un nombre de pestaña NO puede terminar (RT-18).
PARTICULAS = {'y', 'e', 'o', 'u', 'de', 'del', 'la', 'el', 'los', 'las', 'en',
              'con', 'para', 'a', 'al', 'por', 'sin', 'the', 'and', 'of'}


def nombre_pestana(ws, fname):
    """Nombre de pestaña para un checklist del molde A, cuya hoja se llama hoy
    `Sheet` en los 38 ficheros medidos (TEC-28). Sale del título A1 del propio
    libro, no de un diccionario: así vale para las 8 guías sin inventar nada.
    """
    titulo = _norm(ws['A1'].value) or ''
    t = re.sub(r'^checklist\s*(de\s+)?', '', titulo, flags=re.I)
    t = re.split(r'\s+[—–]\s+', t)[0]
    t = RX_ILEGAL_HOJA.sub(' ', t)
    t = re.sub(r'\s+', ' ', t).strip(' ,;.')
    if not t:
        t = re.sub(r'^checklist[-_]', '', fname.rsplit('.', 1)[0])
        t = t.replace('-', ' ').title()
    if len(t) > 31:
        # RT-18/RC-32: `t[:31].rsplit(' ',1)[0]` dejaba «Vajilla, Cristalería y»
        # —cortado EN la conjunción—, que se lee como un fichero roto. Se
        # retrocede palabra a palabra mientras la última sea una partícula.
        palabras = t[:31].split(' ')[:-1]
        while palabras and _norm(palabras[-1]) in PARTICULAS:
            palabras.pop()
        corte = ' '.join(palabras).strip(' ,;.')
        t = corte if len(corte) >= 12 else (t[:30].rstrip(' ,;.') + '…')
    return t.strip() or 'Checklist'


def renombrar_pestana(wb, ws, nuevo, informe=None, fname=''):
    """Renombra sólo si el nombre nuevo es libre y NINGUNA fórmula del libro
    nombra el viejo (un rename silencioso deja `#REF!` en la fórmula)."""
    if ws.title == nuevo:
        return False
    if nuevo in wb.sheetnames:
        return False
    viejo = ws.title
    for hoja in wb.worksheets:
        for row in hoja.iter_rows():
            for c in row:
                if c.data_type == 'f' and isinstance(c.value, str) \
                        and viejo in c.value:
                    if informe is not None:
                        informe.append(
                            fname + ': NO se renombra ' + repr(viejo)
                            + ': lo referencia ' + hoja.title + '!'
                            + c.coordinate)
                    return False
    ws.title = nuevo
    if informe is not None:
        informe.append(fname + ': pestaña ' + repr(viejo) + ' → ' + repr(nuevo)
                       + ' (§1.10/TEC-28)')
    return True


def col_texto(ws):
    """Columna donde vive el texto de `Instrucciones` (A en las 55 medidas)."""
    for col in (1, 2):
        if isinstance(ws.cell(row=4, column=col).value, str):
            return col
    return 1


def linea_instrucciones(ws, texto, rx=None):
    """Sustituye la línea que case con `rx` o la añade al final. Nunca duplica.
    """
    col = col_texto(ws)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return r
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return r
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)
    return destino


def hoja_instrucciones(wb, fname, informe=None, subtitulo=None):
    """Devuelve la hoja `Instrucciones`, **creándola si falta**.

    Medido: 56 de los 111 xlsx no la tienen — los 38 checklists del molde A, los
    15 de panadería (que no la lleva NINGUNO de sus 15 ficheros) y los 3 de
    dark-kitchen (TEC-28). Y es donde viven la línea
    de versión y la bio, que la Fase A no alcanzó a poner en ninguno.

    A la hoja NUEVA sí se le pone el A4 (§1.13 protege las que ya existían, no
    las que no existen: sin `paperSize=9` el censo la contaría como `noprint`).
    """
    if 'Instrucciones' in wb.sheetnames:
        return wb['Instrucciones'], False
    ws0 = wb.worksheets[0]
    titulo = _norm(ws0['A1'].value) or fname.rsplit('.', 1)[0]
    sub = subtitulo or _norm(ws0['A2'].value) or 'AI Chef Pro · aichef.pro'
    ws = wb.create_sheet('Instrucciones', 0)
    wb.active = 0
    ws.column_dimensions['A'].width = 80.0
    val(ws, 'A1', titulo, bold=True)
    ws['A1'].font = Font(bold=True, size=16)
    val(ws, 'A2', sub)
    ws.merge_cells('A1:C1')
    ws.merge_cells('A2:C2')
    val(ws, 'A4', 'Instrucciones de uso', bold=True)
    ws['A4'].font = Font(bold=True, size=12)
    val(ws, 'A5', '1. Rellena las celdas verdes con tus datos; el resto lo '
                  'calcula el libro.')
    val(ws, 'A6', '2. Las hojas están protegidas SIN contraseña para que una '
                  'copia accidental no borre una fórmula.')
    val(ws, 'A8', NOTA_VERDES)
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59,
                                  bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if informe is not None:
        informe.append(fname + ': hoja «Instrucciones» CREADA (§1.10; no la '
                       'tenía, y es donde van versión y bio)')
    return ws, True


def cierre_instrucciones(ws, fname='', con_nota_iva=False, informe=None,
                         pid=None):
    """Bloque de cierre: nota de IVA, «Revisar → Desproteger», BIO y VERSIÓN.

    La bio es una **INSERCIÓN**: no la lleva ninguno de los 111 (medido). Para
    que la 2.ª pasada no la vaya empujando hacia abajo, el bloque se BORRA
    entero primero y se reescribe al final del texto vivo — así su posición
    depende del contenido, no de dónde estaba el bloque antes, aunque un grupo
    haya añadido líneas por el medio.
    """
    col = col_texto(ws)
    patrones = (RX_NOTA_IVA, RX_DESPROTEGER, RX_BIO, RX_VERSION)
    estilo = None
    for r in range(1, ws.max_row + 1):
        cel = ws.cell(row=r, column=col)
        v = cel.value
        if isinstance(v, str) and any(p.search(v) for p in patrones):
            if estilo is None:
                estilo = copy.copy(cel._style)
            cel.value = None
    ultima = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                ultima = r
    lineas = []
    if con_nota_iva:
        lineas.append(NOTA_IVA)
    lineas += [NOTA_DESPROTEGER, BIO_LINE, version_line(pid)]
    fila = ultima + 2
    for texto in lineas:
        cel = ws.cell(row=fila, column=col, value=texto)
        if estilo is not None:
            cel._style = copy.copy(estilo)
        cel.alignment = Alignment(horizontal='left', vertical='top')
        fila += 1
    if informe is not None:
        informe.append(fname + ':' + ws.title + ': bio + versión 2.0'
                       + (' + nota de IVA' if con_nota_iva else '')
                       + ' en A' + str(ultima + 2))
    return fila - 1


# ==========================================================================
# §1.5(b) — la base «sin IVA» dicha en la propia hoja
# ==========================================================================
RX_HOJA_IVA = re.compile(
    r'(p\s*&\s*l|cuenta de resultados|cash\s*flow|flujo de caja|'
    r'ticket medio|escenario|p&l)', re.I)


def hojas_con_nota_iva(wb, fname):
    """Hojas de P&L, cash flow o ticket medio, reconocidas por el NOMBRE DEL
    FICHERO, por el título de la hoja o por su A1 — las tres vías hacen falta:
    en el representante la hoja se llama `Escenarios` y sólo el nombre del
    fichero (`pl-mensual-escenarios.xlsx`) dice que es un P&L."""
    por_fichero = bool(RX_HOJA_IVA.search(fname))
    fuera = []
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        texto = (ws.title or '') + ' ' + str(_norm(ws['A1'].value) or '')
        if por_fichero or RX_HOJA_IVA.search(texto):
            fuera.append(ws)
    return fuera


def nota_iva(wb, fname, informe=None):
    """§1.5(b) — la línea va bajo la cabecera del libro cuando la fila 3 está
    LIBRE. En panadería la fila 3 es la cabecera de la tabla: allí escribirla
    destruiría los encabezados, así que se cae a `Instrucciones` (lo que
    devuelve esta función lo usa `cerrar()` para decidirlo).
    """
    hojas = hojas_con_nota_iva(wb, fname)
    if not hojas:
        return False
    puestas = 0
    for ws in hojas:
        libre = all(ws.cell(row=3, column=c).value is None
                    for c in range(1, ws.max_column + 1))
        actual = _norm(ws['A3'].value)
        if isinstance(actual, str) and RX_NOTA_IVA.match(actual):
            puestas += 1
            continue
        if libre:
            es_caja = bool(RX_HOJA_CAJA.search(
                (ws.title or '') + ' ' + str(_norm(ws['A1'].value) or '')))
            val(ws, 'A3', NOTA_IVA_CAJA if es_caja else NOTA_IVA)
            ws['A3'].font = Font(italic=True, size=9)
            puestas += 1
            if informe is not None:
                informe.append(fname + ':' + ws.title
                               + '!A3: base de IVA declarada (§1.5b)')
    return True if puestas or hojas else False


# ==========================================================================
# §1.11 — metadata
# ==========================================================================
RX_SUFIJO_V = re.compile(r'\s*·\s*v\d+\.\d+\s*$')


def base_subject(wb):
    s = wb.properties.subject
    return RX_SUFIJO_V.sub('', s).strip() if isinstance(s, str) else None


def metadatos(wb, fname, informe=None):
    p = wb.properties
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'
    base = base_subject(wb)
    if base:
        p.subject = base + ' · v2.0'
    if isinstance(p.title, str):
        nuevo = RX_SUFIJO_V.sub('', p.title).strip() + ' · v2.0'
        if p.title != nuevo:
            p.title = nuevo
    if informe is not None:
        informe.append(fname + ': metadata → ' + repr(p.subject) + ' (§1.11)')
    return p.subject


# ==========================================================================
# §1.8 — protección de hoja SIN contraseña
# ==========================================================================
def proteger(ws, informe=None, fname=''):
    """Se desbloquean SÓLO las celdas verdes (y las marcadas por rol).

    Ojo con `password`: `= None` revienta openpyxl y `= ''` escribe el hash de
    la cadena vacía → Excel pediría contraseña justo donde las Instrucciones
    dicen que no hay ninguna. Se deja sin asignar.
    """
    extras = getattr(ws, '_g_editables', set())
    verdes = 0
    for row in ws.iter_rows():
        for c in row:
            if c.__class__.__name__ == 'MergedCell':
                continue
            if (es_verde(c) or c.coordinate in extras) and c.data_type != 'f':
                c.protection = Protection(locked=False)
                verdes += 1
            else:
                c.protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.sort = False
    if informe is not None:
        informe.append(fname + ':' + ws.title + ': protegida sin contraseña ('
                       + str(verdes) + ' celdas editables)')
    return verdes


# ==========================================================================
# Auditoría auxiliar
# ==========================================================================
RX_LITERAL = re.compile(r'(?<![A-Z0-9_$.!])(\d+(?:\.\d+)?)(?![0-9.]*[%)]?\s*'
                        r'[A-Z]*\d*\()')
LITERALES_TOLERADOS = {'0', '1', '2', '3', '4', '5', '12', '30', '100',
                       '0.0', '1.0'}


def literales_sospechosos(wb, fname):
    """Fórmulas con un número clavado dentro que debería ser parámetro en celda.

    No corrige: INFORMA. Los casos citados por la SPEC son
    `escandallo-maestro!H28='=H27/0.28'` (food cost hardcodeado, TEC-21) y
    `escandallo-maestro-<hermano>!G20='=G19*1.10'` (el IVA incrustado, DOM-03).
    """
    fuera = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if c.data_type != 'f' or not isinstance(v, str):
                    continue
                cuerpo = re.sub(r'"[^"]*"', '', v)
                for m in RX_LITERAL.finditer(cuerpo):
                    num = m.group(1)
                    if num in LITERALES_TOLERADOS:
                        continue
                    fuera.append({'fichero': fname, 'hoja': ws.title,
                                  'celda': c.coordinate, 'literal': num,
                                  'formula': v[:90]})
                    break
    return fuera


RX_GUARDA = re.compile(r'^=\s*(IFERROR|IF)\s*\(', re.I)


def contadores(wb, fname):
    r = {'fichero': fname, 'hojas': len(wb.worksheets), 'formulas': 0,
         'verdes': 0, 'cf': 0, 'dv': 0, 'protegidas': 0, 'sin_proteger': 0,
         'divisiones_sin_iferror': 0, 'vacias': 0, 'bio': 0, 'version': 0,
         'a4_ok': 0, 'a4_mal': []}
    for ws in wb.worksheets:
        r['cf'] += sum(len(cf.rules) for cf in ws.conditional_formatting)
        r['dv'] += len(ws.data_validations.dataValidation)
        if ws.protection.sheet:
            r['protegidas'] += 1
        else:
            r['sin_proteger'] += 1
        fit = (ws.sheet_properties.pageSetUpPr.fitToPage
               if ws.sheet_properties.pageSetUpPr else None)
        pie = (ws.oddFooter.center.text
               if ws.oddFooter and ws.oddFooter.center else None)
        es_texto = ws.title in ('Instrucciones', 'Índice', 'Indice')
        if ws.page_setup.paperSize == 9 and (es_texto or (fit and pie)):
            r['a4_ok'] += 1
        else:
            r['a4_mal'].append(ws.title)
        for row in ws.iter_rows():
            for c in row:
                if es_verde(c):
                    r['verdes'] += 1
                v = c.value
                if c.data_type == 'f' and isinstance(v, str):
                    r['formulas'] += 1
                    if '/' in v and not RX_GUARDA.match(v):
                        r['divisiones_sin_iferror'] += 1
                elif v == '':
                    r['vacias'] += 1
                elif isinstance(v, str):
                    if RX_BIO.search(v):
                        r['bio'] += 1
                    if RX_VERSION.match(v):
                        r['version'] += 1
    return r


# ==========================================================================
# CICLO: aplicar (antes de los grupos) · cerrar (después)
# ==========================================================================
def detectar(wb, fname, pid=''):
    """§1.1 — TODO lo que el motor necesita saber del libro, ANTES de escribir.

    Lanza `MoldeDesconocido` si un checklist no encaja: el aborto es la
    característica, no el efecto colateral (§7-bis.11).
    """
    info = {'fichero': fname, 'hojas': wb.sheetnames,
            'variante_pl': variante_pl(wb) if 'pl-mensual' in fname else None,
            'checklists': []}
    if es_checklist(fname):
        for ws in wb.worksheets:
            if ws.title == 'Instrucciones':
                continue
            molde, fila = molde_checklist(ws, fname, pid)
            info['checklists'].append({'hoja': ws.title, 'molde': molde,
                                       'fila_cabecera': fila})
    return info


def aplicar(wb, fname, informe, pid=None):
    """§1 transversal ANTES del trabajo de los grupos.

    Aquí va sólo lo que NO depende de las filas que los grupos van a insertar:
    metadata, la hoja de Instrucciones, el nombre de la pestaña y la nota de
    IVA. Todo lo que se mide (última fila de datos, categorías, formatos) va en
    `cerrar()`.
    """
    pid = pid or CTX.get('producto')
    info = detectar(wb, fname, pid or '')
    metadatos(wb, fname, informe)
    hoja_instrucciones(wb, fname, informe)
    if es_checklist(fname):
        for ch in info['checklists']:
            ws = wb[ch['hoja']]
            if ws.title == 'Sheet':
                nuevo = nombre_pestana(ws, fname)
                if renombrar_pestana(wb, ws, nuevo, informe, fname):
                    ch['hoja'] = nuevo
    nota_iva(wb, fname, informe)
    wb._g_info = info
    return info


def cerrar(wb, fname, informe, pid=None, proteger_hojas=True):
    """Cierre común, DESPUÉS de los grupos: si se protegiera antes, cada celda
    que un grupo creara nacería bloqueada aunque fuese verde.
    """
    pid = pid or CTX.get('producto')
    info = getattr(wb, '_g_info', None) or detectar(wb, fname, pid or '')
    resumen = []
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        # la cabecera puede haberse movido si un grupo insertó filas arriba
        fila_cab = None
        molde = None
        if es_checklist(fname):
            molde, fila_cab = molde_checklist(ws, fname, pid or '')
        else:
            fila_cab = fila_cabecera_tabla(ws)
        if fila_cab:
            formato_por_etiqueta(ws, fila_cab, informe, fname)
        elif not molde:
            formato_ficha(ws, informe, fname)
        if molde:
            r = cerrar_checklist(ws, molde, fila_cab, fname, informe, pid)
            if r:
                resumen.append(dict(r, hoja=ws.title))
        arreglar_cabecera_escandallo(ws, informe, fname)
        ensanchar_etiquetas(ws, informe, fname)
        retirar_verde_de_calculadas(ws, informe, fname)
        verdes_por_dv(ws, informe, fname)

    # el censo cuenta `''` como defecto `empty_str`
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value == '' and c.__class__.__name__ != 'MergedCell':
                    c.value = None

    lleva_iva = bool(hojas_con_nota_iva(wb, fname))
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            cierre_instrucciones(ws, fname, lleva_iva, informe, pid)
            if proteger_hojas:
                proteger(ws, informe, fname)
            continue
        validaciones(ws, informe, fname)
        if proteger_hojas:
            proteger(ws, informe, fname)
    metadatos(wb, fname, None)
    info['checklists_cerrados'] = resumen
    return info
