#!/usr/bin/env python3
"""
grupo_b.py — Grupo B de la familia «Guías Cómo Montar» v2.0: **checklists y
cronograma** (`guias-v2-SPEC.md` §3 entero).

Ficheros que toca (MEDIDO el 2026-08-29 abriendo los 8 productos): los **8**
checklists del representante, los **6** de cada uno de los 5 hermanos, los **6**
de panadería (molde B) y los **2** de dark-kitchen (moldes C y D), más
`cronograma-apertura-gantt.xlsx` (7 guías; dark-kitchen no lo lleva).

Es el grupo donde vive la **exposición jurídica** del producto: un checklist
legal caducado no es un defecto cosmético, es una instrucción errónea que el
cliente ejecuta (§3 cabecera).

QUÉ HACE ESTE MÓDULO (lógica de FAMILIA) y qué hace `contenido_<pid>/b.py`
-------------------------------------------------------------------------
Aquí no hay ni una fila de contenido: sólo el motor de aplicación. Las filas,
los textos, los importes y los parámetros de cada guía viven en
`contenido_<pid>/b.py`, con la fuente de cada cifra anotada. Este módulo:

  §3.1/§3.2/§3.4  `sustituir` / `insertar` / `renombrar_categoria` /
                  `anotar` sobre los checklists, **por molde y por CABECERA**,
                  nunca por letra de columna fija (los hermanos tienen otras
                  cabeceras y otros nombres de fichero:
                  `checklist-diseno-sala-mexicana.xlsx`,
                  `checklist-equipamiento-cocina-japonesa.xlsx`).
  §3.3            lo que los moldes NO traen y el motor no cubre: el semáforo
                  de desviación con `ISNUMBER` (molde C) y la **reparación de
                  los totales/contadores nativos** cuando la inserción de filas
                  mueve el final de los datos (`F40='=SUM(F5:F39)'` de
                  dark-kitchen y los `COUNTIF(A4:A34,"✓")` de panadería: openpyxl
                  **no** reajusta los rangos al insertar).
  §3.4            la columna calculada «Menú degustación» de la vajilla y la
                  fila de «instalación, transporte y puesta en marcha (%)» del
                  equipamiento, las dos **con el parámetro en celda**.
  §3.5            el Gantt: cabeceras de mes a NÚMERO, columnas `Mes inicio`,
                  `Duración (meses)`, `Depende de` y `Días`, **barra por formato
                  condicional** y aviso cuando el plan se sale del horizonte.

TRES MOLDES DE GANTT, no uno (medido, y ninguno lo dice la SPEC)
---------------------------------------------------------------
El §3.5 describe el Gantt del representante como si fuera el de las 7 guías. No
lo es, y aplicarle su rejilla a un hermano escribiría encima de la columna de
tareas:

  · **G1 — representante**: hojas `Instrucciones` + `Gantt`; cabecera en la
    fila 4 con `A='Fase / Tarea'`, `B='Responsable'`, `C='Estado'`,
    `D='Inicio'`, `E='Fin'` y `F4:W4 = M1…M18`. Las filas de fase van solas en
    la columna A (`FASE 1: PLANIFICACIÓN`). **La rejilla está vacía.**
  · **G2 — los 5 hermanos**: cabecera en la fila 4 con `A='#'`, `B='Fase'`,
    `C='Tarea'` y `D4:O4 = M1…M12`. **No hay Responsable, ni Estado, ni
    Inicio/Fin**, y la rejilla YA trae marcas `'X'`.
  · **G3 — panadería**: hoja única `Gantt apertura`, cabecera en la fila 3 con
    `A='Hito'`, `B='Responsable'`, `C3:H3 = M1…M6` e `I='Notas críticas'`.
    Marcas `'■'`.

Por eso la rejilla se localiza **buscando la tira de cabeceras `M<n>`**, y los
roles (tarea, responsable, estado, inicio, fin) se resuelven leyendo la
cabecera: lo que no existe, no se usa. Donde ya hay marcas, el `Mes inicio` y la
`Duración` **se DEDUCEN de ellas** (es una medición, no una invención) y las
marcas se retiran, porque si se quedaran, al cambiar el `Mes inicio` la barra
nueva se pintaría en un sitio y la `X` vieja seguiría en otro.

IDEMPOTENCIA (main.py la comprueba con una 2.ª pasada sobre un clon)
--------------------------------------------------------------------
Todo lo que se escribe es **absoluto y detectable**: las inserciones se
deduplican por el TEXTO de la tarea, las sustituciones dejan de encontrar su
patrón una vez aplicadas, las columnas nuevas del Gantt se reconocen por su
cabecera y las fórmulas se re-emiten con los índices MEDIDOS en esa pasada. Se
escribe con `motor.f()` para que `main.py` verifique una a una que quedaron
cacheadas.

REGLA DE CIFRAS (§7-bis / regla capital): ningún importe del sector se teclea.
Si el contenido no trae el importe con fuente, la celda de coste **queda vacía**
—nunca `0`, que en un presupuesto se lee como «gratis»— con la nota de que hay
que pedir presupuesto. La celda ya es verde y editable: eso ES la
parametrización.

⚠️ Los ficheros de la familia llevan ESPACIO FINO (U+202F) y GUION NO SEPARABLE
(U+2011). Aquí se referencian por escape (`motor.NARROW`, `motor.NOBRK`), nunca
escribiendo el carácter.
"""
import copy
import glob
import os
import re

from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import column_index_from_string, get_column_letter

import motor

AQUI = os.path.dirname(os.path.abspath(__file__))
DL = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(AQUI))),
                  'astro-site', 'public', 'dl')

SPEC = 'guias-v2-SPEC.md §3'

#: Color de la BARRA del Gantt. Deliberadamente **fuera** de la paleta de
#: semáforos (verde/ámbar/rojo): una barra de planificación no es un juicio
#: sobre el dato, y pintarla de verde haría leer «bien» a un mes que sólo está
#: planificado.
BARRA_BG = 'D9E1F2'
BARRA_FG = '1F3864'

RX_MES = re.compile(r'^M\s*(\d{1,2})$', re.I)
FMT_MES = '"M"0'

#: Cabeceras de las columnas que añade el Gantt (§3.5). El orden es fijo: es lo
#: que hace que la 2.ª pasada las encuentre donde las dejó la primera.
COL_MES_INICIO = 'Mes inicio'
COL_DURACION = 'Duración (meses)'
COL_DEPENDE = 'Depende de'
COL_DIAS = 'Días'

NOTA_GANTT = ('1. Rellena «' + COL_MES_INICIO + '» y «' + COL_DURACION
              + '»: la barra de los meses se pinta sola a partir de esas dos '
                'celdas (formato condicional).')
RX_NOTA_GANTT_VIEJA = re.compile(
    r'^\s*1\.\s*(Cada fase tiene una barra|Fases y tareas)', re.I)
NOTA_GANTT_2 = ('2. «' + COL_DEPENDE + '» dice qué tarea tiene que estar '
                'terminada antes; «' + COL_DIAS + '» sale de Inicio y Fin.')
RX_NOTA_GANTT_2_VIEJA = re.compile(
    r'^\s*2\.\s*(Rellena las fechas|Marca las celdas)', re.I)

#: Etiqueta del bloque de parámetros que este grupo deja en `Instrucciones`.
CAB_PARAMETROS = 'Parámetros de este libro (edítalos: el libro recalcula)'


# ==========================================================================
# Utilidades de estructura (todo por CABECERA, nunca por letra fija)
# ==========================================================================
def _txt(v):
    return v.strip() if isinstance(v, str) else v


def _norm(v):
    """Texto normalizado para comparar tareas: sin dobles espacios ni tipografía
    fina. Los `.md`/`.xlsx` de la familia llevan U+202F y U+2011 y un mismo
    ítem puede venir escrito de las dos formas."""
    if not isinstance(v, str):
        return ''
    t = v.replace(motor.NARROW, ' ').replace(motor.NOBRK, '-')
    t = t.replace(' ', ' ')
    return re.sub(r'\s+', ' ', t).strip().lower()


def _col(letra):
    return column_index_from_string(letra)


def _roles(ws, fila_cab, molde=None):
    """Mapa rol → letra de columna, leído de la CABECERA.

    Los cuatro moldes escriben lo mismo con nombres distintos: `Tarea / Ítem`
    (A), `Punto de control` / `Trámite` / `Equipo / utensilio` / `Acción` /
    `Punto a verificar` (B, panadería), `Equipo/Concepto` (C, dark-kitchen) y
    `Tarea` (D). Resolverlo por letra rompería en cuanto cambia el molde.
    """
    cab = motor.cabeceras(ws, fila_cab)
    r = {'num': None, 'categoria': None, 'tarea': None, 'responsable': None,
         'estado': None, 'coste': None, 'notas': None, 'desviacion': None,
         'real': None}
    for L, v in cab.items():
        t = _norm(v)
        if not t:
            continue
        if t == '#':
            r['num'] = L
        elif t in ('categoría', 'categoria', 'zona', 'fase'):
            r['categoria'] = L
        elif t in ('responsable', 'proveedor'):
            r['responsable'] = L
        elif t == 'estado':
            r['estado'] = L
        elif t in ('coste est. (€)', 'coste estimado (€)', 'presupuesto (€)'):
            r['coste'] = L
        elif t == 'real (€)':
            r['real'] = L
        elif t == 'desviación (%)':
            r['desviacion'] = L
        elif t in ('notas', 'nota', 'observaciones', 'notas críticas'):
            r['notas'] = L
        elif r['tarea'] is None and re.search(
                r'tarea|ítem|item|equipo|concepto|punto de control|'
                r'punto a verificar|trámite|tramite|acción|accion|hito', t):
            r['tarea'] = L
    if molde == 'B':
        # El molde se DEFINE por la casilla de la columna A, así que la columna
        # de texto es la siguiente. No es una posición adivinada: es la firma
        # estructural del molde (motor.molde_checklist).
        r['tarea'] = 'B'
        r['categoria'] = None
    if r['tarea'] is None:
        r['tarea'] = 'C' if molde in ('A', 'C', 'D') else 'B'
    return r


def _fin_datos(ws, molde, fila_cab, rol_tarea):
    """Última fila de DATOS, MEDIDA (§9, gate de recuento: los recuentos del R1
    no cuadran entre sí — su inventario dice «checklist-legal 42 filas» y lo
    medido son 40)."""
    fin = fila_cab
    for r in range(fila_cab + 1, ws.max_row + 1):
        if molde == 'B':
            if _txt(ws.cell(row=r, column=1).value) in ('☐', '✓', 'N/A'):
                fin = r
            elif ws.cell(row=r, column=2).value is not None:
                fin = r                      # fila de sección
        else:
            v = ws.cell(row=r, column=1).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                fin = r
    if fin == fila_cab and rol_tarea:        # molde sin numeración en A
        for r in range(fila_cab + 1, ws.max_row + 1):
            if isinstance(ws.cell(row=r,
                                  column=_col(rol_tarea)).value, str):
                fin = r
    return fin


def _estilo_de(ws, fila_modelo, columna):
    return copy.copy(ws.cell(row=fila_modelo, column=columna)._style)


def _desmontar_merges(ws, desde):
    """Quita las combinaciones que empiezan en `desde` o más abajo.

    `insert_rows` de openpyxl **no** mueve los rangos combinados: el pie
    `A46:H46` se quedaría en la fila 46 mientras su texto baja a la 54, y el
    fichero abriría con una banda combinada en mitad de los datos. El pie lo
    vuelve a escribir y a combinar `motor.cerrar_checklist()` al final.
    """
    fuera = []
    for m in [str(x) for x in ws.merged_cells.ranges]:
        if int(re.search(r'(\d+)', m.split(':')[0]).group(1)) >= desde:
            fuera.append(m)
    for m in fuera:
        ws.unmerge_cells(m)
    return fuera


# ==========================================================================
# §3.1/§3.2/§3.4 — sustituir, anotar e insertar filas en un checklist
# ==========================================================================
def _buscar_fila(ws, rol_tarea, texto, fila_cab, fin):
    objetivo = _norm(texto)
    col = _col(rol_tarea)
    for r in range(fila_cab + 1, fin + 1):
        if _norm(ws.cell(row=r, column=col).value) == objetivo:
            return r
    # segunda vuelta: coincidencia por prefijo, para tolerar que el original
    # lleve un paréntesis de más. Nunca por «contiene»: «Registro de
    # temperaturas de cocción» contiene a «Registro de temperaturas».
    for r in range(fila_cab + 1, fin + 1):
        actual = _norm(ws.cell(row=r, column=col).value)
        if actual and (actual.startswith(objetivo)
                       or objetivo.startswith(actual)) \
                and abs(len(actual) - len(objetivo)) <= 12:
            return r
    return None


def _escribir_celda(ws, fila, letra, valor, fmt=None, verde=False,
                    modelo=None):
    if not letra:
        return
    cel = ws.cell(row=fila, column=_col(letra))
    if modelo is not None:
        cel._style = _estilo_de(ws, modelo, _col(letra))
    if valor is not None:
        cel.value = valor
    if fmt:
        cel.number_format = fmt
    if verde:
        cel.fill = PatternFill('solid', fgColor=motor.VERDE)
        cel.protection = Protection(locked=False)


def _sustituir(ws, roles, fila_cab, fin, regla, cambios, fname):
    """§3.1 — cambia el TEXTO (y, si la regla lo dice, responsable/coste/nota)
    de una fila existente. No borra la fila: la sustituye, que es lo que pide la
    SPEC («se sustituye, no se borra dejando hueco»).
    """
    fila = _buscar_fila(ws, roles['tarea'], regla['buscar'], fila_cab, fin)
    if fila is None:
        # ¿ya aplicada? Entonces el texto NUEVO está donde estaba el viejo.
        if _buscar_fila(ws, roles['tarea'], regla['tarea'], fila_cab,
                        fin) is not None:
            return 'ya'
        return None
    _escribir_celda(ws, fila, roles['tarea'], regla['tarea'])
    if regla.get('responsable') and roles['responsable']:
        _escribir_celda(ws, fila, roles['responsable'], regla['responsable'])
    if 'coste' in regla and roles['coste']:
        cel = ws.cell(row=fila, column=_col(roles['coste']))
        cel.value = regla['coste']          # puede ser None: «sin dato» = vacío
        cel.number_format = motor.FMT_EUR
    if regla.get('notas') and roles['notas']:
        _escribir_celda(ws, fila, roles['notas'], regla['notas'])
    cambios.append(fname + ':' + ws.title + '!' + roles['tarea'] + str(fila)
                   + ': «' + str(regla['buscar'])[:44] + '» → «'
                   + str(regla['tarea'])[:60] + '» [' + regla.get('id', '')
                   + ' · ' + regla.get('fuente', 'SPEC') + ']')
    return fila


def _anotar(ws, roles, fila_cab, fin, regla, cambios, fname):
    """Añade/actualiza SÓLO la nota de una fila que ya está bien redactada."""
    if not roles['notas']:
        return None
    fila = _buscar_fila(ws, roles['tarea'], regla['buscar'], fila_cab, fin)
    if fila is None:
        return None
    cel = ws.cell(row=fila, column=_col(roles['notas']))
    if _norm(cel.value) == _norm(regla['notas']):
        return fila
    cel.value = regla['notas']
    cambios.append(fname + ':' + ws.title + '!' + roles['notas'] + str(fila)
                   + ': nota [' + regla.get('id', '') + ']')
    return fila


def _renombrar_categoria(ws, roles, fila_cab, fin, viejo, nuevo, cambios,
                         fname):
    """COM-34 — «Reputación» pasa a «Prensa y notoriedad (NO influye en la
    inspección)»: contigua a ítems de estrella sugería lo que el propio cap. 17
    niega. Se renombra la categoría, no se borra: son acciones legítimas."""
    if not roles['categoria']:
        return 0
    col = _col(roles['categoria'])
    n = 0
    for r in range(fila_cab + 1, fin + 1):
        if _norm(ws.cell(row=r, column=col).value) == _norm(viejo):
            ws.cell(row=r, column=col).value = nuevo
            n += 1
    if n:
        cambios.append(fname + ':' + ws.title + ': categoría «' + viejo
                       + '» → «' + nuevo + '» en ' + str(n) + ' filas [COM-34]')
    return n


def _fila_destino(ws, roles, fila_cab, fin, nueva, molde):
    """Dónde entra la fila nueva: **detrás de la última de su categoría** (o de
    su sección, en el molde B), y al final si la categoría no existe todavía.

    Insertar al final sin más rompería el agrupamiento por categoría del que
    viven los subtotales `SUMIF` que escribe el motor (§1.9). Cuando no hay
    columna de categoría —el Gantt del representante no la tiene: las fases van
    en filas sueltas de la columna A— el contenido ancla con `despues_de`, que
    nombra la tarea tras la que va la nueva.
    """
    ancla = nueva.get('despues_de')
    if ancla:
        fila = _buscar_fila(ws, roles['tarea'], ancla, fila_cab, fin)
        if fila:
            return fila + 1
    if molde == 'B':
        seccion = nueva.get('seccion')
        if seccion:
            ultimo = None
            dentro = False
            for r in range(fila_cab + 1, fin + 1):
                a = _txt(ws.cell(row=r, column=1).value)
                b = ws.cell(row=r, column=2).value
                es_seccion = a not in ('☐', '✓', 'N/A') and b is None \
                    and isinstance(a, str)
                if es_seccion:
                    dentro = _norm(a) == _norm(seccion)
                    if dentro:
                        ultimo = r
                elif dentro:
                    ultimo = r
            if ultimo:
                return ultimo + 1
        return fin + 1
    cat = nueva.get('categoria')
    if cat and roles['categoria']:
        col = _col(roles['categoria'])
        ultimo = None
        for r in range(fila_cab + 1, fin + 1):
            if _norm(ws.cell(row=r, column=col).value) == _norm(cat):
                ultimo = r
        if ultimo:
            return ultimo + 1
    return fin + 1


def _insertar(ws, roles, fila_cab, fin, nueva, molde, cambios, fname):
    """§3.2/§3.4 — inserta UNA fila nueva. Devuelve el nuevo `fin`.

    Deduplica por el TEXTO de la tarea: en la 2.ª pasada la fila ya está y no se
    duplica (idempotencia). Y **la deduplicación es de familia, no del
    representante**: `checklist-appcc` de la guía japonesa ya trae una categoría
    `Anisakis (PCC)` con sus controles, así que las filas de congelación
    preventiva que el §3.2 manda añadir NO se le añaden dos veces.
    """
    if _buscar_fila(ws, roles['tarea'], nueva['tarea'], fila_cab, fin) \
            is not None:
        return fin, 'ya'
    for alias in nueva.get('equivale_a', ()):
        if _buscar_fila(ws, roles['tarea'], alias, fila_cab, fin) is not None:
            return fin, 'equivalente'
    destino = _fila_destino(ws, roles, fila_cab, fin, nueva, molde)
    modelo = destino - 1 if destino - 1 > fila_cab else fila_cab + 1
    _desmontar_merges(ws, destino)
    ws.insert_rows(destino, 1)
    if modelo >= destino:
        modelo += 1
    for c in range(1, ws.max_column + 1):
        ws.cell(row=destino, column=c)._style = _estilo_de(ws, modelo, c)
        ws.cell(row=destino, column=c).value = None
    if molde == 'B':
        _escribir_celda(ws, destino, 'A', '☐')
    _escribir_celda(ws, destino, roles['tarea'], nueva['tarea'])
    if roles['categoria'] and nueva.get('categoria'):
        _escribir_celda(ws, destino, roles['categoria'], nueva['categoria'])
    if roles['responsable'] and nueva.get('responsable'):
        _escribir_celda(ws, destino, roles['responsable'],
                        nueva['responsable'])
    if roles['estado'] and nueva.get('estado'):
        _escribir_celda(ws, destino, roles['estado'], nueva['estado'])
    if roles['coste']:
        cel = ws.cell(row=destino, column=_col(roles['coste']))
        # «Sin dato» se escribe vacío, NUNCA 0 (§7-bis.13): un 0 en una columna
        # de presupuesto se lee «esto no cuesta nada» y falsea el TOTAL.
        cel.value = nueva.get('coste')
        cel.number_format = motor.FMT_EUR
    for extra, valor in (nueva.get('columnas') or {}).items():
        letra = roles.get(extra)
        if letra:
            _escribir_celda(ws, destino, letra, valor)
    if roles['notas'] and nueva.get('notas'):
        _escribir_celda(ws, destino, roles['notas'], nueva['notas'])
    cambios.append(fname + ':' + ws.title + '!' + str(destino) + ': + «'
                   + str(nueva['tarea'])[:66] + '» ['
                   + nueva.get('id', '') + ' · '
                   + nueva.get('fuente', 'SPEC') + ']')
    return fin + 1, destino


def _renumerar(ws, roles, fila_cab, fin, molde):
    """La columna `#` vuelve a ser 1..n después de insertar (moldes A, C y D).
    Sin esto, un checklist con filas nuevas enseña `… 32, 33, 33, 34`."""
    if molde == 'B' or not roles['num']:
        return 0
    col = _col(roles['num'])
    n = 0
    for r in range(fila_cab + 1, fin + 1):
        n += 1
        ws.cell(row=r, column=col).value = n
    return n


# ==========================================================================
# §3.3 — lo que los moldes YA traen y hay que reparar tras insertar
# ==========================================================================
RX_SUM = re.compile(r'^=SUM\(([A-Z]{1,2})(\d+):([A-Z]{1,2})(\d+)\)$', re.I)
RX_COUNTIF = re.compile(
    r'^=COUNTIF\(([A-Z]{1,2})(\d+):([A-Z]{1,2})(\d+),(.+)\)$', re.I)


def _reparar_rangos_nativos(ws, fila_cab, fin, cambios, fname):
    """openpyxl **no** reajusta los rangos de las fórmulas al insertar filas.

    Dos casos MEDIDOS en la familia, los dos por debajo de los datos:
      · dark-kitchen `checklist-apertura-legal!F40='=SUM(F5:F39)'` y
        `checklist-equipamiento-obra!E45/F45='=SUM(E5:E44)'` — el TOTAL nativo
        del molde, que el motor respeta y no duplica (§3.3);
      · panadería `checklist-legal!C36='=COUNTIF(A4:A34,"✓")'` y
        `E36='=COUNTIF(B4:B34,"?*")'` — el contador propio del molde B, que la
        SPEC manda **respetar**.
    Si se insertan filas y no se reajustan, el TOTAL deja de sumar las últimas
    partidas y el contador deja de contarlas: el fichero sigue abriendo, con el
    build en verde, dando un número equivocado.
    """
    tocadas = 0
    for r in range(fin + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if cel.data_type != 'f' or not isinstance(cel.value, str):
                continue
            m = RX_SUM.match(cel.value.strip())
            if m and m.group(1).upper() == m.group(3).upper():
                nueva = ('=SUM(' + m.group(1) + str(fila_cab + 1) + ':'
                         + m.group(3) + str(fin) + ')')
                if nueva != cel.value:
                    motor.f(ws, cel.coordinate, nueva,
                            fmt=cel.number_format)
                    tocadas += 1
                    cambios.append(fname + ':' + ws.title + '!'
                                   + cel.coordinate + ': TOTAL nativo '
                                   + m.group(0) + ' → ' + nueva + ' (§3.3)')
                continue
            m = RX_COUNTIF.match(cel.value.strip())
            if m and m.group(1).upper() == m.group(3).upper():
                # El contador del molde B arranca en la fila de la CABECERA
                # (A4), no en la primera casilla: cuenta también las filas de
                # sección. Se conserva ese arranque tal cual estaba.
                nueva = ('=COUNTIF(' + m.group(1) + m.group(2) + ':'
                         + m.group(3) + str(fin) + ',' + m.group(5) + ')')
                if nueva != cel.value:
                    motor.f(ws, cel.coordinate, nueva,
                            fmt=cel.number_format)
                    tocadas += 1
                    cambios.append(fname + ':' + ws.title + '!'
                                   + cel.coordinate + ': contador propio '
                                   + m.group(0)[:34] + '… → …' + str(fin)
                                   + ' (§3.3/§7-bis.17)')
    return tocadas


def _semaforo_desviacion(ws, roles, fila_cab, fin, umbral_pct, cambios,
                         fname):
    """§3.3 molde C — la desviación de CAPEX en rojo por encima del umbral, con
    la guarda `ISNUMBER` (§1.6).

    `G5='=IF(E5=0,"",((F5-E5)/E5))'` devuelve **texto vacío** mientras no haya
    presupuesto: sin `ISNUMBER`, Excel evalúa `""<0` como FALSO pero
    `"">umbral` también como FALSO por casualidad, y en cuanto la celda trae un
    aviso de texto el semáforo se enciende en la celda que dice que no hay dato.
    El umbral va **en celda**, no dentro de la fórmula.
    """
    if not roles['desviacion']:
        return None
    col = roles['desviacion']
    rango = col + str(fila_cab + 1) + ':' + col + str(fin)
    # El umbral vive en la fila 3 (libre en los cuatro moldes: 1 título,
    # 2 subtítulo, 3 vacía, 4 cabecera) y en la MISMA hoja: el formato
    # condicional con referencia a otra hoja no es portable entre visores.
    etiqueta = 'Umbral de desviación que se marca en rojo (%)'
    celda_umbral = None
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(row=3, column=c).value) == _norm(etiqueta):
            celda_umbral = get_column_letter(c + 1) + '3'
    if celda_umbral is None:
        col_et = roles['tarea'] or 'C'
        col_val = get_column_letter(_col(col_et) + 1)
        motor.val(ws, col_et + '3', etiqueta)
        motor.val(ws, col_val + '3', umbral_pct, fmt=motor.FMT_PCT,
                  verde_=True)
        celda_umbral = col_val + '3'
        motor.dv_porcentaje(ws, [celda_umbral])
    else:
        ws[celda_umbral].value = umbral_pct
        ws[celda_umbral].number_format = motor.FMT_PCT
    ancla = '$' + col + str(fila_cab + 1)
    motor.semaforo_isnumber(ws, rango, ancla, operador='>',
                            umbral='$' + celda_umbral[0] + '$3')
    cambios.append(fname + ':' + ws.title + ': semáforo de desviación sobre '
                   + rango + ' con ISNUMBER y umbral en ' + celda_umbral
                   + ' (§3.3/§1.6)')
    return celda_umbral


# ==========================================================================
# Parámetros en `Instrucciones` (§1.5: parámetro en celda, nunca literal)
# ==========================================================================
def _param(wb, clave_etiqueta, valor, fmt, nota, cambios, fname,
           porcentaje=False):
    """Deja `etiqueta | valor | nota` en la hoja `Instrucciones` y devuelve la
    referencia absoluta de la celda del valor (`Instrucciones!$B$12`).

    Se ancla por la ETIQUETA: en la 2.ª pasada la encuentra y reutiliza la misma
    fila, así que la posición no se desplaza (que es como se rompe la
    idempotencia en este paquete). El bloque de cierre del motor (bio, versión,
    nota de IVA) se reescribe siempre DESPUÉS del último contenido vivo, así que
    convive con esto sin pisarlo.
    """
    ws = wb['Instrucciones']
    col = motor.col_texto(ws)
    fila = None
    for r in range(1, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=col).value) == _norm(clave_etiqueta):
            fila = r
            break
    if fila is None:
        ultima = 0
        cabecera = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value is not None:
                    ultima = r
            if _norm(ws.cell(row=r, column=col).value) == _norm(CAB_PARAMETROS):
                cabecera = r
        if cabecera is None:
            motor.val(ws, get_column_letter(col) + str(ultima + 2),
                      CAB_PARAMETROS, bold=True)
            ultima += 2
        fila = ultima + 1
        cambios.append(fname + ':Instrucciones!A' + str(fila)
                       + ': parámetro «' + clave_etiqueta[:52]
                       + '» en celda (§1.5)')
    letra_val = get_column_letter(col + 1)
    letra_nota = get_column_letter(col + 2)
    motor.val(ws, get_column_letter(col) + str(fila), clave_etiqueta)
    motor.val(ws, letra_val + str(fila), valor, fmt=fmt, verde_=True)
    motor.val(ws, letra_nota + str(fila), nota, wrap=True)
    ws.column_dimensions[letra_val].width = 14.0
    ws.column_dimensions[letra_nota].width = 62.0
    if porcentaje:
        motor.dv_porcentaje(ws, [letra_val + str(fila)])
    else:
        motor.dv_numerica(ws, [letra_val + str(fila)], minimo=0)
    return "Instrucciones!$" + letra_val + '$' + str(fila)


# ==========================================================================
# §3.4 — la fila «instalación, transporte y puesta en marcha (%)»
# ==========================================================================
def _fila_porcentual(ws, wb, roles, fila_cab, fin, cfg, cambios, fname):
    """DOM-20 — «Instalación, transporte y puesta en marcha» como PORCENTAJE
    calculado sobre el resto de partidas, no como un importe inventado.

    La fórmula suma el rango **partido en dos tramos alrededor de su propia
    fila**: `SUM(G5:G61)+SUM(G63:G90)`. Referirse al rango entero y restarse a
    sí misma sería una autorreferencia circular — exactamente el bug
    `Break-Even!B12='=B12*(1-B8)'` que la SPEC documenta como reintroducido por
    5 de los 7 generadores.
    """
    if not roles['coste']:
        return None
    fila = _buscar_fila(ws, roles['tarea'], cfg['tarea'], fila_cab, fin)
    if fila is None:
        return None
    ref = _param(wb, cfg['etiqueta'], cfg['valor'], motor.FMT_PCT,
                 cfg['nota'], cambios, fname, porcentaje=True)
    col = roles['coste']
    tramos = []
    if fila - 1 >= fila_cab + 1:
        tramos.append('SUM($' + col + '$' + str(fila_cab + 1) + ':$' + col
                      + '$' + str(fila - 1) + ')')
    if fin >= fila + 1:
        tramos.append('SUM($' + col + '$' + str(fila + 1) + ':$' + col + '$'
                      + str(fin) + ')')
    if not tramos:
        return None
    formula = motor.iferror('ROUND((' + '+'.join(tramos) + ')*' + ref + ',2)')
    motor.f(ws, col + str(fila), formula, fmt=motor.FMT_EUR)
    cambios.append(fname + ':' + ws.title + '!' + col + str(fila)
                   + ': instalación y puesta en marcha = ' + formula[:80]
                   + ' [DOM-20 · §3.4]')
    return col + str(fila)


# ==========================================================================
# §3.4 — la segunda columna «Menú degustación» de la vajilla (DOM-21)
# ==========================================================================
def _columna_calculada(ws, wb, roles, fila_cab, fin, cfg, cambios, fname):
    """DOM-21 — no se redimensiona la dotación de carta (que es correcta para su
    supuesto): se añade una **segunda columna** con la regla explícita
    `piezas = plazas × uds/comensal × rotación de lavado`, para que el cliente
    vea las dos y elija (§7.1).

    Dos columnas nuevas: la de ENTRADA (`uds/comensal`, verde) y la CALCULADA.
    Donde el cliente no ha puesto factor, la calculada devuelve `""`, **nunca
    `0`**: un 0 en «copas necesarias» se lee como «no necesitas ninguna».
    """
    col_entrada, col_calc = None, None
    # ⚠️ Hasta `max_column`, NUNCA más allá: `ws.cell()` **crea** la celda, y
    # mirar una columna de más hacía crecer `max_column` en cada pasada. El
    # motor combina el pie hasta `max_column`, así que la idempotencia saltaba
    # con «cambia merges (A70:K70 → A70:L70)» — una columna fantasma por pasada.
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row=fila_cab, column=c).value)
        if v == _norm(cfg['cabecera_entrada']):
            col_entrada = get_column_letter(c)
        elif v == _norm(cfg['cabecera_calculada']):
            col_calc = get_column_letter(c)
    if col_entrada is None or col_calc is None:
        base = ws.max_column
        col_entrada = get_column_letter(base + 1)
        col_calc = get_column_letter(base + 2)
        motor.val(ws, col_entrada + str(fila_cab), cfg['cabecera_entrada'],
                  bold=True, wrap=True)
        motor.val(ws, col_calc + str(fila_cab), cfg['cabecera_calculada'],
                  bold=True, wrap=True)
        for L in (col_entrada, col_calc):
            ws.cell(row=fila_cab, column=_col(L))._style = _estilo_de(
                ws, fila_cab, _col(roles['coste'] or 'B'))
            ws.column_dimensions[L].width = 16.0
        cambios.append(fname + ':' + ws.title + ': columnas «'
                       + cfg['cabecera_entrada'] + '» y «'
                       + cfg['cabecera_calculada'] + '» en ' + col_entrada
                       + '/' + col_calc + ' [DOM-21 · §3.4]')
    refs = []
    for p in cfg['parametros']:
        refs.append(_param(wb, p['etiqueta'], p['valor'],
                           p.get('formato', motor.FMT_ENT), p['nota'],
                           cambios, fname))
    precarga = dict((_norm(k), v) for k, v in
                    (cfg.get('precarga') or {}).items())
    puestas = 0
    for r in range(fila_cab + 1, fin + 1):
        tarea = _norm(ws.cell(row=r, column=_col(roles['tarea'])).value)
        cel_e = ws.cell(row=r, column=_col(col_entrada))
        if tarea in precarga and cel_e.value is None:
            cel_e.value = precarga[tarea]
            puestas += 1
        cel_e.number_format = '#,##0.0'
        cel_e.fill = PatternFill('solid', fgColor=motor.VERDE)
        cel_e.protection = Protection(locked=False)
        formula = ('=IF($' + col_entrada + str(r) + '="","",ROUND($'
                   + col_entrada + str(r) + '*' + '*'.join(refs) + ',0))')
        motor.f(ws, col_calc + str(r), formula, fmt=motor.FMT_ENT)
    if puestas:
        cambios.append(fname + ':' + ws.title + ': ' + str(puestas)
                       + ' factores uds/comensal precargados (el resto los '
                         'pone el cliente; la columna calculada da "" sin '
                         'factor)')
    return col_entrada, col_calc


# ==========================================================================
# §3.5 — el Gantt
# ==========================================================================
def _rejilla_meses(ws):
    """Localiza la tira de cabeceras `M1…Mn`. Devuelve
    `(fila_cabecera, primera_col, ultima_col, [numeros])` o `None`.

    Se busca la tira, no una columna concreta: es lo único que comparten los
    tres moldes de Gantt (F:W en el representante, D:O en los hermanos, C:H en
    panadería). Reconoce tanto el texto original `'M1'` como la cabecera ya
    convertida a número con formato `"M"0` (2.ª pasada).
    """
    for fila in (4, 3, 5, 2):
        run = []
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=fila, column=c)
            num = None
            m = RX_MES.match(_txt(cel.value) or '') if isinstance(
                cel.value, str) else None
            if m:
                num = int(m.group(1))
            elif isinstance(cel.value, int) and not isinstance(cel.value, bool) \
                    and '"M"' in (cel.number_format or ''):
                num = cel.value
            if num is not None:
                run.append((c, num))
            elif run:
                break
        if len(run) >= 3:
            return fila, run[0][0], run[-1][0], [n for _, n in run]
    return None


def _cabeceras_mes_a_numero(ws, fila, ini, fin_col, nums, cambios, fname):
    """`'M1'` (texto) → `1` con formato `"M"0`.

    No es cosmética: la barra se pinta con `F$4>=$X6`, y en Excel **cualquier
    texto es mayor que cualquier número**, así que comparando contra la cadena
    `'M1'` la regla se cumpliría en todas las columnas y la barra pintaría la
    fila entera. La celda sigue viéndose «M1».
    """
    n = 0
    for (c, num) in zip(range(ini, fin_col + 1), nums):
        cel = ws.cell(row=fila, column=c)
        if isinstance(cel.value, str):
            cel.value = num
            n += 1
        cel.number_format = FMT_MES
        cel.alignment = Alignment(horizontal='center')
    if n:
        cambios.append(fname + ':' + ws.title + ': ' + str(n) + ' cabeceras '
                       'de mes de texto a número con formato "M"0 (§3.5: sin '
                       'esto la comparación de la barra es texto vs número y '
                       'pinta la fila entera)')
    return n


def _columnas_gantt(ws, fila_cab, ultima_mes, con_fechas, cambios, fname):
    """Las columnas nuevas van **detrás** de la rejilla de meses.

    `insert_cols` habría sido más bonito visualmente, pero openpyxl no mueve ni
    los anchos de columna, ni los rangos combinados, ni el formato condicional,
    ni las validaciones: la banda de título `A1:V1` se quedaría a mitad y la
    rejilla perdería sus anchos de 5,0. Detrás, y por cabecera, es reversible y
    idempotente.
    """
    quiero = [COL_MES_INICIO, COL_DURACION, COL_DEPENDE]
    if con_fechas:
        quiero.append(COL_DIAS)
    mapa = {}
    for c in range(1, ws.max_column + 1):
        v = _txt(ws.cell(row=fila_cab, column=c).value)
        if v in quiero:
            mapa[v] = get_column_letter(c)
    siguiente = max(ultima_mes, ws.max_column) + 1
    for etiqueta in quiero:
        if etiqueta in mapa:
            continue
        L = get_column_letter(siguiente)
        cel = ws.cell(row=fila_cab, column=siguiente, value=etiqueta)
        cel._style = _estilo_de(ws, fila_cab, ultima_mes)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
        ws.column_dimensions[L].width = 13.0 if etiqueta != COL_DEPENDE else 30.0
        mapa[etiqueta] = L
        siguiente += 1
        cambios.append(fname + ':' + ws.title + '!' + L + str(fila_cab)
                       + ': columna «' + etiqueta + '» (§3.5)')
    return mapa


def _es_fila_tarea(ws, fila, rol_tarea, roles_extra):
    """Distingue una TAREA de una cabecera de fase.

    En el representante la fase va sola en la columna A (`FASE 1:
    PLANIFICACIÓN`, sin Estado); en los hermanos hay un `#` numérico en A y la
    fase va en su propia columna B. Se decide por estructura, no por el texto.
    """
    v = ws.cell(row=fila, column=_col(rol_tarea)).value
    if not isinstance(v, str) or not v.strip():
        return False
    if roles_extra.get('num'):
        a = ws.cell(row=fila, column=_col(roles_extra['num'])).value
        return isinstance(a, (int, float)) and not isinstance(a, bool)
    acompanantes = [roles_extra.get(k) for k in ('estado', 'responsable')
                    if roles_extra.get(k)]
    if acompanantes:
        return any(ws.cell(row=fila, column=_col(L)).value is not None
                   for L in acompanantes)
    return not re.match(r'^(fase|bloque)\s*\d', v.strip(), re.I)


def _deducir_de_marcas(ws, fila, ini, fin_col, nums):
    """`Mes inicio` y `Duración` DEDUCIDOS de las marcas que el fichero ya trae
    (`'X'` en los 5 hermanos, `'■'` en panadería). Es una MEDICIÓN del propio
    entregable, no una cifra inventada.

    Devuelve `(mes, duracion, huecos)`. `huecos` avisa de una tarea marcada en
    meses no consecutivos: la barra por formato condicional es un tramo
    continuo, así que ahí se pierde información y hay que decirlo.
    """
    marcados = []
    for (c, num) in zip(range(ini, fin_col + 1), nums):
        v = ws.cell(row=fila, column=c).value
        if v is not None and str(v).strip() != '':
            marcados.append(num)
    if not marcados:
        return None, None, 0
    primero, ultimo = min(marcados), max(marcados)
    huecos = (ultimo - primero + 1) - len(marcados)
    return primero, ultimo - primero + 1, huecos


def _limpiar_marcas(ws, fila, ini, fin_col):
    n = 0
    for c in range(ini, fin_col + 1):
        cel = ws.cell(row=fila, column=c)
        if cel.value is not None:
            cel.value = None
            n += 1
    return n


def _gantt(wb, fname, cambios, contenido):
    """§3.5 completo sobre los tres moldes de Gantt."""
    cfg = (getattr(contenido, 'GANTT', None) or {}) if contenido else {}
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        rejilla = _rejilla_meses(ws)
        if not rejilla:
            continue
        fila_cab, ini, fin_col, nums = rejilla
        roles = _roles(ws, fila_cab)
        rol_tarea = roles['tarea'] or 'A'
        _cabeceras_mes_a_numero(ws, fila_cab, ini, fin_col, nums, cambios,
                                fname)

        # --- tareas nuevas (§3.5: arrendamiento y financiación) -----------
        fin = _fin_datos(ws, None, fila_cab, rol_tarea)
        for nueva in cfg.get('nuevas_tareas', ()):
            fin, _ = _insertar(ws, roles, fila_cab, fin, nueva, None, cambios,
                               fname)
        for regla in cfg.get('sustituciones', ()):
            _sustituir(ws, roles, fila_cab, fin, regla, cambios, fname)
        _renumerar(ws, roles, fila_cab, fin, None)

        # `Días` sólo tiene sentido donde el molde trae fechas: G1 las tiene
        # (D=Inicio, E=Fin) y G2/G3 no. Se comprueba, no se supone.
        tiene_inicio = tiene_fin = None
        for c in range(1, ws.max_column + 1):
            t = _norm(ws.cell(row=fila_cab, column=c).value)
            if t == 'inicio':
                tiene_inicio = get_column_letter(c)
            elif t == 'fin':
                tiene_fin = get_column_letter(c)
        con_fechas = bool(tiene_inicio and tiene_fin)

        mapa = _columnas_gantt(ws, fila_cab, fin_col, con_fechas, cambios,
                               fname)
        col_ini, col_dur = mapa[COL_MES_INICIO], mapa[COL_DURACION]
        col_dep = mapa[COL_DEPENDE]

        # --- precarga: contenido de la guía > marcas medidas --------------
        plan = dict((_norm(k), v) for k, v in
                    (cfg.get('tareas') or {}).items())
        deducidas, huecos_total, primera_fila = 0, 0, None
        for r in range(fila_cab + 1, fin + 1):
            if not _es_fila_tarea(ws, r, rol_tarea, roles):
                continue
            if primera_fila is None:
                primera_fila = r
            clave = _norm(ws.cell(row=r, column=_col(rol_tarea)).value)
            datos = plan.get(clave)
            mes = dur = dep = None
            if datos:
                mes, dur = datos.get('mes'), datos.get('duracion')
                dep = datos.get('depende')
            if mes is None:
                mes, dur, huecos = _deducir_de_marcas(ws, r, ini, fin_col,
                                                      nums)
                if mes is not None:
                    deducidas += 1
                    huecos_total += huecos
            if mes is not None:
                _limpiar_marcas(ws, r, ini, fin_col)
            cel_m = ws.cell(row=r, column=_col(col_ini))
            cel_d = ws.cell(row=r, column=_col(col_dur))
            if cel_m.value is None and mes is not None:
                cel_m.value = mes
            if cel_d.value is None and dur is not None:
                cel_d.value = dur
            for cel, fmt in ((cel_m, motor.FMT_ENT), (cel_d, motor.FMT_ENT)):
                cel.number_format = fmt
                cel.fill = PatternFill('solid', fgColor=motor.VERDE)
                cel.protection = Protection(locked=False)
                cel.alignment = Alignment(horizontal='center')
            cel_dep = ws.cell(row=r, column=_col(col_dep))
            if cel_dep.value is None and dep:
                cel_dep.value = dep
            cel_dep.fill = PatternFill('solid', fgColor=motor.VERDE)
            cel_dep.protection = Protection(locked=False)
            cel_dep.alignment = Alignment(horizontal='left', wrap_text=True)
            if con_fechas:
                col_dias = mapa[COL_DIAS]
                motor.f(ws, col_dias + str(r),
                        '=IFERROR(IF(OR($' + tiene_inicio + str(r) + '="",$'
                        + tiene_fin + str(r) + '=""),"",$' + tiene_fin + str(r)
                        + '-$' + tiene_inicio + str(r) + '),"")',
                        fmt=motor.FMT_ENT)
        if deducidas:
            cambios.append(
                fname + ':' + ws.title + ': ' + str(deducidas) + ' tareas con '
                'Mes inicio/Duración DEDUCIDOS de las marcas que ya traía el '
                'fichero (medición, no invención); marcas retiradas para que '
                'la barra siga al dato' + (' · ' + str(huecos_total)
                                           + ' meses sueltos perdidos al '
                                             'convertir marcas no '
                                             'consecutivas en un tramo'
                                           if huecos_total else ''))

        # --- fechas: formato dd/mm/yyyy + DV (TEC-15) ---------------------
        if con_fechas and primera_fila:
            coords = []
            for L in (tiene_inicio, tiene_fin):
                for r in range(primera_fila, fin + 1):
                    if _es_fila_tarea(ws, r, rol_tarea, roles):
                        cel = ws.cell(row=r, column=_col(L))
                        cel.number_format = motor.FMT_FECHA
                        cel.fill = PatternFill('solid', fgColor=motor.VERDE)
                        cel.protection = Protection(locked=False)
                        coords.append(L + str(r))
            motor.dv_fecha(ws, coords)
            cambios.append(fname + ':' + ws.title + ': Inicio/Fin en '
                           'dd/mm/yyyy con DV de fecha y columna «'
                           + COL_DIAS + '» (TEC-15)')

        # --- la barra, por formato condicional (TEC-14, DOM-37, COM-19) ---
        if primera_fila:
            rejilla_rango = (get_column_letter(ini) + str(primera_fila) + ':'
                             + get_column_letter(fin_col) + str(fin))
            ancla_mes = get_column_letter(ini) + '$' + str(fila_cab)
            formula = ('=AND(ISNUMBER($' + col_ini + str(primera_fila)
                       + '),ISNUMBER($' + col_dur + str(primera_fila) + '),'
                       + ancla_mes + '>=$' + col_ini + str(primera_fila) + ','
                       + ancla_mes + '<$' + col_ini + str(primera_fila) + '+$'
                       + col_dur + str(primera_fila) + ')')
            motor.regla_expresion(ws, rejilla_rango, formula, bg=BARRA_BG,
                                  fg=BARRA_FG, parar=True)
            cambios.append(fname + ':' + ws.title + ': barra del Gantt por '
                           'formato condicional sobre ' + rejilla_rango
                           + ' → ' + formula + ' (TEC-14/DOM-37/COM-19)')
            # Aviso: el plan no cabe en el horizonte del propio Gantt. El
            # horizonte NO se teclea: sale de la última cabecera de mes.
            tope = ('MAX($' + get_column_letter(ini) + '$' + str(fila_cab)
                    + ':$' + get_column_letter(fin_col) + '$' + str(fila_cab)
                    + ')')
            aviso = ('=AND(ISNUMBER($' + col_ini + str(primera_fila)
                     + '),ISNUMBER($' + col_dur + str(primera_fila) + '),$'
                     + col_ini + str(primera_fila) + '+$' + col_dur
                     + str(primera_fila) + '-1>' + tope + ')')
            motor.regla_expresion(
                ws, col_ini + str(primera_fila) + ':' + col_dur + str(fin),
                aviso, bg=motor.CF_ROJO_BG, fg=motor.CF_ROJO_FG, parar=True)
            cambios.append(fname + ':' + ws.title + ': aviso en rojo cuando '
                           'Mes inicio + Duración se sale del horizonte del '
                           'propio Gantt (' + aviso[:70] + '…)')

        # --- Estado: desplegable y semáforo, si el molde lo tiene ---------
        if roles['estado'] and primera_fila:
            coords = [roles['estado'] + str(r)
                      for r in range(primera_fila, fin + 1)
                      if _es_fila_tarea(ws, r, rol_tarea, roles)]
            opciones = motor.opciones_dv(ws, coords[0]) if coords else []
            if opciones:
                motor.dv_lista(ws, coords, opciones)
                vocab = [t for t in motor.SEM_ESTADO if t[0] in opciones]
                if vocab:
                    motor.semaforo_texto(
                        ws, roles['estado'] + str(primera_fila) + ':'
                        + roles['estado'] + str(fin), tuple(vocab))

    # --- las Instrucciones dejan de prometer lo que no había --------------
    if 'Instrucciones' in wb.sheetnames:
        ws = wb['Instrucciones']
        motor.linea_instrucciones(ws, NOTA_GANTT, RX_NOTA_GANTT_VIEJA)
        motor.linea_instrucciones(ws, NOTA_GANTT_2, RX_NOTA_GANTT_2_VIEJA)
        cambios.append(fname + ':Instrucciones: la promesa «cada fase tiene '
                       'una barra de duración estimada» pasa a describir cómo '
                       'se pinta (§3.5)')


# ==========================================================================
# Checklists
# ==========================================================================
def _checklist(wb, fname, cambios, contenido, cfg):
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        molde, fila_cab = motor.molde_checklist(ws, fname,
                                                motor.CTX.get('producto') or '')
        roles = _roles(ws, fila_cab, molde)
        fin = _fin_datos(ws, molde, fila_cab, roles['tarea'])
        antes = fin - fila_cab

        for viejo, nuevo in (cfg.get('renombrar_categoria') or {}).items():
            _renombrar_categoria(ws, roles, fila_cab, fin, viejo, nuevo,
                                 cambios, fname)
        for regla in cfg.get('sustituciones', ()):
            if _sustituir(ws, roles, fila_cab, fin, regla, cambios,
                          fname) is None:
                cambios.append(fname + ':' + ws.title + ': AVISO — no se '
                               'encuentra la fila a sustituir «'
                               + str(regla['buscar'])[:60] + '» ['
                               + regla.get('id', '') + ']')
        for regla in cfg.get('notas', ()):
            _anotar(ws, roles, fila_cab, fin, regla, cambios, fname)
        for nueva in cfg.get('nuevas', ()):
            fin, _ = _insertar(ws, roles, fila_cab, fin, nueva, molde, cambios,
                               fname)
        _renumerar(ws, roles, fila_cab, fin, molde)
        if fin != fila_cab + antes:
            _reparar_rangos_nativos(ws, fila_cab, fin, cambios, fname)

        if cfg.get('columna_calculada'):
            _columna_calculada(ws, wb, roles, fila_cab, fin,
                               cfg['columna_calculada'], cambios, fname)
        if cfg.get('fila_porcentual'):
            _fila_porcentual(ws, wb, roles, fila_cab, fin,
                             cfg['fila_porcentual'], cambios, fname)
        if roles['desviacion']:
            _semaforo_desviacion(ws, roles, fila_cab, fin,
                                 cfg.get('umbral_desviacion', 0.10), cambios,
                                 fname)
        anuncia = cfg.get('anuncia')
        items = fin - fila_cab
        cambios.append(fname + ':' + ws.title + ': molde ' + molde + ' · '
                       + str(antes) + ' → ' + str(items) + ' ítems MEDIDOS'
                       + (' (la tarjeta anuncia ' + str(anuncia) + ')'
                          if anuncia else ''))
        if anuncia and items < anuncia:
            cambios.append(fname + ':' + ws.title + ': AVISO — quedan '
                           + str(anuncia - items) + ' ítems por debajo de lo '
                           'anunciado (§3.4/COM-17/COM-18)')


# ==========================================================================
# Contrato con main.py
# ==========================================================================
def ficheros(ctx):
    """Los checklists y el cronograma **de ese producto**, leídos del disco.

    No es una lista fija: los hermanos renombran los ficheros
    (`checklist-diseno-sala-mexicana.xlsx`,
    `checklist-equipamiento-cocina-japonesa.xlsx`,
    `checklist-salida-humos.xlsx`, `checklist-apertura-legal.xlsx`) y una lista
    escrita a mano dejaría fuera justo los que cambian de nombre.
    """
    pid = (ctx or {}).get('producto') or motor.CTX.get('producto')
    if not pid:
        return []
    carpeta = os.path.join(DL, pid)
    fuera = []
    for p in sorted(glob.glob(os.path.join(carpeta, '*.xlsx'))):
        n = os.path.basename(p)
        if n.startswith('checklist-') or n.startswith('cronograma-'):
            fuera.append(n)
    return fuera


PROPIOS = []          # el §1 del motor se aplica entero a todos estos ficheros


def post(wb, fname, cambios, registro, contenido):
    """Se ejecuta DESPUÉS de `motor.aplicar()` (que ya ha creado la hoja
    `Instrucciones` donde este grupo deja sus parámetros) y ANTES de
    `motor.cerrar()` (que pondrá verdes, DV, formatos, el bloque de TOTAL/%
    completado y la protección sobre lo que aquí se escriba)."""
    conf = (getattr(contenido, 'CHECKLISTS', None) or {}) if contenido else {}
    if fname.startswith('cronograma-'):
        _gantt(wb, fname, cambios, contenido)
        return
    if motor.es_checklist(fname):
        _checklist(wb, fname, cambios, contenido, conf.get(fname, {}))


# ==========================================================================
# Demostraciones propias (pycel) — sobre COPIAS, nunca sobre los entregables
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    import contextlib
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__


def _copia(carpeta, fname, destino):
    import shutil
    os.makedirs(destino, exist_ok=True)
    fuera = os.path.join(destino, 'b-' + fname)
    shutil.copy2(os.path.join(carpeta, fname), fuera)
    return fuera


def demos(carpeta, origen, destino, contenido):
    """Cada cálculo NUEVO de este grupo, evaluado con pycel **cambiando los
    inputs**: un número que sale bien una vez puede ser una constante.

    Se comprueban, por este orden:
      1. la barra del Gantt (se evalúa la fórmula de la regla de formato
         condicional en una celda auxiliar, mes a mes, antes y después de mover
         `Mes inicio`: la barra tiene que desplazarse el mismo número de meses);
      2. `Días` = Fin − Inicio, y `""` cuando falta una de las dos fechas
         (con el libro en blanco no puede salir un `0` que se lea como «esta
         tarea dura cero días»);
      3. la fila de instalación y puesta en marcha: al subir una partida de
         equipamiento 1.000 €, sube su 12 % y el TOTAL sube 1.120 €;
      4. la columna «Menú degustación»: al doblar las plazas se dobla la
         dotación, y sin factor por comensal devuelve `""`, no `0`.
    """
    resultado = {'grupo_b': {}, 'fallos': []}
    fallos = resultado['fallos']
    detalle = resultado['grupo_b']
    cfgs = (getattr(contenido, 'CHECKLISTS', None) or {}) if contenido else {}
    gcfg = (getattr(contenido, 'GANTT', None) or {}) if contenido else {}

    # ---- 1 y 2: el Gantt -------------------------------------------------
    import openpyxl
    for fname in sorted(os.listdir(carpeta)):
        if not fname.startswith('cronograma-') or not fname.endswith('.xlsx'):
            continue
        path = os.path.join(carpeta, fname)
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            if ws.title == 'Instrucciones':
                continue
            rejilla = _rejilla_meses(ws)
            if not rejilla:
                continue
            fila_cab, ini, fin_col, nums = rejilla
            mapa = {}
            for c in range(1, ws.max_column + 1):
                v = _txt(ws.cell(row=fila_cab, column=c).value)
                if v in (COL_MES_INICIO, COL_DURACION, COL_DIAS):
                    mapa[v] = get_column_letter(c)
            if COL_MES_INICIO not in mapa:
                fallos.append(fname + ':' + ws.title + ': sin columna «'
                              + COL_MES_INICIO + '» (§3.5 sin aplicar)')
                continue
            roles = _roles(ws, fila_cab)
            rol_tarea = roles['tarea'] or 'A'
            fila = None
            for r in range(fila_cab + 1, ws.max_row + 1):
                if _es_fila_tarea(ws, r, rol_tarea, roles) and \
                        isinstance(ws.cell(row=r,
                                           column=_col(mapa[COL_MES_INICIO])
                                           ).value, (int, float)):
                    fila = r
                    break
            if fila is None:
                fallos.append(fname + ':' + ws.title + ': ninguna tarea con '
                              'Mes inicio precargado — la barra no se pintaría '
                              'en ninguna fila')
                continue
            copia = _copia(carpeta, fname, destino)
            wb2 = openpyxl.load_workbook(copia)
            ws2 = wb2[ws.title]
            libre = max(ws2.max_column, fin_col) + 6
            ci, cd = mapa[COL_MES_INICIO], mapa[COL_DURACION]
            for k, (c, num) in enumerate(zip(range(ini, fin_col + 1), nums)):
                ws2.cell(row=fila, column=libre + k).value = (
                    '=IF(AND(ISNUMBER($' + ci + str(fila) + '),ISNUMBER($'
                    + cd + str(fila) + '),' + get_column_letter(c) + '$'
                    + str(fila_cab) + '>=$' + ci + str(fila) + ','
                    + get_column_letter(c) + '$' + str(fila_cab) + '<$' + ci
                    + str(fila) + '+$' + cd + str(fila) + '),1,0)')
            wb2.save(copia)
            xl = _pycel(copia)
            pref = "'" + ws.title + "'!"
            antes = [_ev(xl, pref + get_column_letter(libre + k) + str(fila))
                     for k in range(len(nums))]
            mes0 = ws.cell(row=fila, column=_col(ci)).value
            _ev(xl, pref + ci + str(fila))
            xl.set_value(pref + ci + str(fila), mes0 + 2)
            despues = [_ev(xl, pref + get_column_letter(libre + k) + str(fila))
                       for k in range(len(nums))]
            pintados_antes = [nums[i] for i, v in enumerate(antes) if v == 1]
            pintados_despues = [nums[i] for i, v in enumerate(despues)
                                if v == 1]
            ok = (pintados_antes and pintados_despues
                  and min(pintados_despues) - min(pintados_antes) == 2
                  and len(pintados_antes) == len(pintados_despues))
            detalle.setdefault('gantt_barra', []).append({
                'fichero': fname, 'hoja': ws.title, 'fila': fila,
                'tarea': ws.cell(row=fila, column=_col(rol_tarea)).value,
                'celda_mes_inicio': ci + str(fila),
                'meses_pintados': pintados_antes,
                'mes_inicio_+2': pintados_despues, 'ok': bool(ok)})
            if not ok:
                fallos.append(fname + ':' + ws.title + '!' + ci + str(fila)
                              + ': la barra NO sigue a «' + COL_MES_INICIO
                              + '» (' + str(pintados_antes) + ' → '
                              + str(pintados_despues) + ')')
            # ---- 2: la columna Días ------------------------------------
            if COL_DIAS in mapa:
                cdias = mapa[COL_DIAS] + str(fila)
                vacio = _ev(xl, pref + cdias)
                col_i = col_f = None
                for c in range(1, ws.max_column + 1):
                    t = _norm(ws.cell(row=fila_cab, column=c).value)
                    if t == 'inicio':
                        col_i = get_column_letter(c)
                    elif t == 'fin':
                        col_f = get_column_letter(c)
                xl.set_value(pref + col_i + str(fila), 46000)
                xl.set_value(pref + col_f + str(fila), 46030)
                lleno = _ev(xl, pref + cdias)
                detalle.setdefault('gantt_dias', []).append({
                    'fichero': fname, 'celda': cdias,
                    'libro_en_blanco': vacio, 'con_fechas_46000_46030': lleno,
                    'ok': vacio == '' and lleno == 30})
                if vacio != '':
                    fallos.append(fname + ':' + cdias + ': con el libro en '
                                  'blanco «' + COL_DIAS + '» devuelve '
                                  + repr(vacio) + ' en vez de "" (§7-bis.13)')
                if lleno != 30:
                    fallos.append(fname + ':' + cdias + ': Fin−Inicio da '
                                  + repr(lleno) + ' y debía dar 30')

    # ---- 3: la fila porcentual de instalación ---------------------------
    for fname, cfg in cfgs.items():
        fp = cfg.get('fila_porcentual')
        if not fp or not os.path.isfile(os.path.join(carpeta, fname)):
            continue
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wb.worksheets:
            if ws.title == 'Instrucciones':
                continue
            molde, fila_cab = motor.molde_checklist(ws, fname, '')
            roles = _roles(ws, fila_cab, molde)
            fin = _fin_datos(ws, molde, fila_cab, roles['tarea'])
            fila = _buscar_fila(ws, roles['tarea'], fp['tarea'], fila_cab, fin)
            if fila is None:
                fallos.append(fname + ': no existe la fila «'
                              + fp['tarea'][:50] + '» (DOM-20)')
                continue
            col = roles['coste']
            coord = col + str(fila)
            # la fila del TOTAL que escribe el motor (§1.9)
            total = None
            for r in range(fin + 1, ws.max_row + 1):
                if _norm(ws.cell(row=r, column=1).value) == _norm(
                        motor.ETIQUETA_TOTAL):
                    total = col + str(r)
            copia = _copia(carpeta, fname, destino)
            xl = _pycel(copia)
            pref = "'" + ws.title + "'!"
            v0 = _ev(xl, pref + coord)
            t0 = t1 = None
            t0 = _ev(xl, pref + total) if total else None
            # sube 1.000 € una partida cualquiera de equipamiento
            objetivo = None
            for r in range(fila_cab + 1, fin + 1):
                if r == fila:
                    continue
                if isinstance(ws.cell(row=r, column=_col(col)).value,
                              (int, float)):
                    objetivo = col + str(r)
                    break
            if objetivo is None:
                fallos.append(fname + ': ninguna partida con importe que mover '
                              'para demostrar el % de instalación')
                continue
            base = _ev(xl, pref + objetivo)
            xl.set_value(pref + objetivo, base + 1000)
            v1 = _ev(xl, pref + coord)
            t1 = _ev(xl, pref + total) if total else None
            pct = fp['valor']
            ok = (isinstance(v0, (int, float)) and isinstance(v1, (int, float))
                  and abs((v1 - v0) - 1000 * pct) < 0.02)
            ok_total = (t0 is None
                        or (isinstance(t0, (int, float))
                            and isinstance(t1, (int, float))
                            and abs((t1 - t0) - 1000 * (1 + pct)) < 0.02))
            detalle['instalacion_pct'] = {
                'fichero': fname, 'celda': coord, 'valor': v0,
                'celda_movida': objetivo, 'valor_tras_+1000': v1,
                'delta_esperado': round(1000 * pct, 2),
                'total': total, 'total_antes': t0, 'total_despues': t1,
                'ok': bool(ok and ok_total)}
            if not ok:
                fallos.append(fname + ':' + coord + ': la instalación no sigue '
                              'al equipamiento (' + repr(v0) + ' → '
                              + repr(v1) + ', esperado +'
                              + str(round(1000 * pct, 2)) + ')')
            if not ok_total:
                fallos.append(fname + ':' + str(total) + ': el TOTAL no '
                              'recoge la partida + su instalación ('
                              + repr(t0) + ' → ' + repr(t1) + ')')

    # ---- 4: la columna calculada de la vajilla --------------------------
    for fname, cfg in cfgs.items():
        cc = cfg.get('columna_calculada')
        if not cc or not os.path.isfile(os.path.join(carpeta, fname)):
            continue
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wb.worksheets:
            if ws.title == 'Instrucciones':
                continue
            molde, fila_cab = motor.molde_checklist(ws, fname, '')
            roles = _roles(ws, fila_cab, molde)
            fin = _fin_datos(ws, molde, fila_cab, roles['tarea'])
            col_e = col_c = None
            for c in range(1, ws.max_column + 1):
                v = _norm(ws.cell(row=fila_cab, column=c).value)
                if v == _norm(cc['cabecera_entrada']):
                    col_e = get_column_letter(c)
                elif v == _norm(cc['cabecera_calculada']):
                    col_c = get_column_letter(c)
            if not col_e or not col_c:
                fallos.append(fname + ': faltan las columnas de menú '
                              'degustación (DOM-21)')
                continue
            con = sin = None
            for r in range(fila_cab + 1, fin + 1):
                if isinstance(ws.cell(row=r, column=_col(col_e)).value,
                              (int, float)) and con is None:
                    con = r
                elif ws.cell(row=r, column=_col(col_e)).value is None \
                        and sin is None:
                    sin = r
            copia = _copia(carpeta, fname, destino)
            xl = _pycel(copia)
            pref = "'" + ws.title + "'!"
            v0 = _ev(xl, pref + col_c + str(con)) if con else None
            vacio = _ev(xl, pref + col_c + str(sin)) if sin else ''
            ref_plazas = cc['parametros'][0]
            wbi = openpyxl.load_workbook(copia)
            wsi = wbi['Instrucciones']
            celda_plazas = None
            for r in range(1, wsi.max_row + 1):
                if _norm(wsi.cell(row=r, column=1).value) == _norm(
                        ref_plazas['etiqueta']):
                    celda_plazas = 'B' + str(r)
            base = _ev(xl, "'Instrucciones'!" + celda_plazas)
            xl.set_value("'Instrucciones'!" + celda_plazas, base * 2)
            v1 = _ev(xl, pref + col_c + str(con)) if con else None
            ok = (isinstance(v0, (int, float)) and isinstance(v1, (int, float))
                  and abs(v1 - 2 * v0) <= 1 and vacio == '')
            detalle['menu_degustacion'] = {
                'fichero': fname, 'celda': col_c + str(con), 'valor': v0,
                'plazas': celda_plazas, 'plazas_x2': v1,
                'sin_factor_por_comensal': repr(vacio), 'ok': bool(ok)}
            if not isinstance(v0, (int, float)):
                fallos.append(fname + ':' + col_c + str(con) + ': la columna '
                              '«' + cc['cabecera_calculada'] + '» no evalúa ('
                              + repr(v0) + ')')
            elif abs(v1 - 2 * v0) > 1:
                fallos.append(fname + ':' + col_c + str(con) + ': doblar las '
                              'plazas no dobla la dotación (' + repr(v0)
                              + ' → ' + repr(v1) + ')')
            if vacio != '':
                fallos.append(fname + ':' + col_c + str(sin) + ': sin factor '
                              'por comensal devuelve ' + repr(vacio)
                              + ' en vez de "" (§7-bis.13)')

    # ---- 5: recuento de ítems contra lo que anuncia la tarjeta ----------
    recuentos = []
    for fname, cfg in cfgs.items():
        path = os.path.join(carpeta, fname)
        if not cfg.get('anuncia') or not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            if ws.title == 'Instrucciones':
                continue
            molde, fila_cab = motor.molde_checklist(ws, fname, '')
            roles = _roles(ws, fila_cab, molde)
            fin = _fin_datos(ws, molde, fila_cab, roles['tarea'])
            items = fin - fila_cab
            recuentos.append({'fichero': fname, 'items_medidos': items,
                              'anuncia': cfg['anuncia'],
                              'ok': items >= cfg['anuncia']})
            if items < cfg['anuncia']:
                fallos.append(fname + ': ' + str(items) + ' ítems medidos < '
                              + str(cfg['anuncia']) + ' anunciados (§3.4 / '
                              'gate de recuento del §9)')
    detalle['recuento_items'] = recuentos
    detalle['spec'] = SPEC
    return resultado
