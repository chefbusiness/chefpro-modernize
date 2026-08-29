#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 de la familia «Planes de Negocio».

    python3 main.py --producto <pid> [--dry-run] [--solo motor|a|b|c] \
                    [--json informe.json]

`--producto` es el id de carpeta de `astro-site/public/dl/` (ojo: dos de los
diez NO siguen el patrón `plan-negocio-*`: son `plan-catering-tematico-eventos`
y `plan-chef-privado-showcooking-eventos`; los ids exactos los manda
`netlify/functions/get-download-urls.ts:673-760`).

SEGURIDAD (§7-bis.8 y regla dura del encargo)
---------------------------------------------
Sin `--dry-run` este script ABORTA salvo que el orquestador exporte
`PLANES_APPLY=1`, y aun entonces hace **respaldo previo** de la carpeta del
producto en el scratchpad antes de tocar un solo byte. En `--dry-run` trabaja
sobre una copia en
`$CLAUDE_SCRATCHPAD/dryrun/<pid>` y `astro-site/public/dl/` no se abre para
escritura en ningún momento.

QUÉ HACE, EN ORDEN
------------------
1. Copia (dry-run) o respalda (apply) la carpeta del producto.
2. Por cada `.xlsx`: `motor.detectar()` → **aborta con exit 2 si un fichero no
   casa con ninguna firma de la familia** (§1.1); `pre()` de los grupos →
   `motor.aplicar()` (§1 antes) → `post()` de los grupos → `motor.cerrar()`
   (§1 después) → `wb.save()`.
3. Idempotencia: 2.ª pasada sobre un CLON y `diff` de huellas. 0 diferencias o
   es fallo.
4. `inject_cache.py` (siempre el último paso que toca el zip).
5. Verificación `data_only` de **cada fórmula nueva**: las que devuelven `None`
   sin ser `""` por diseño son fallo. Y encima el gate
   **`blancos_contaminados`** (CRIT-01, 0 exigido), que es el que de verdad
   los distingue: caché `None` +
   CUERPO sin ningún literal `""` propio + guarda FALSA = la celda quedó muda
   porque un operando valía `""`, no porque nadie lo quisiera.
6. Gates de §9 medibles sobre xlsx: no latinos, ortografía, formatos
   (0 importes con %, 0 recuentos con €), referencias colgando, 0 xlsx con 0
   fórmulas, ítems de checklist, A4, bio y versión.
7. `censo-entregables.py --only <carpeta> --fail --quiet`.
8. Demostraciones con pycel (§2.11/§3.9): anualidad sin `PMT`, TIR/VAN/payback
   sin `IRR`, `COUNTIF` en vez de `COUNTA`, guarda `IFERROR` viva, protección,
   bio+versión y **detección de los 30 moldes de la familia** (lectura de
   `dl/`, sin escribir).
9. Informe JSON con TODOS los hallazgos (nunca recortados) y `exit` 0/1/2.

CONTRATO DE LOS GRUPOS (`grupo_a.py`, `grupo_b.py`, `grupo_c.py`)
-----------------------------------------------------------------
Módulos hermanos de este fichero. Ninguno existe todavía: `main.py` funciona
con los que encuentre y anota como fallo un grupo **pedido** que no carga (un
paquete a medias no puede declararse terminado).

    FICHEROS   lista de nombres de fichero que el grupo toca, o
               `ficheros(det_por_fichero, contenido)` si depende del producto
    PROPIOS    ficheros que el grupo construye ENTEROS (main NO les aplica el
               §1 transversal: los reescribe el grupo)
    pre(...)   antes del §1 — para leer el estado original
    post(...)  después del §1 y antes del cierre — el trabajo del grupo
    demos(carpeta, origen, demos_dir) -> dict con clave 'fallos'

`pre` y `post` se invocan por NOMBRE de parámetro (se les pasa sólo lo que
declaren), así que las firmas válidas son cualquier subconjunto de:

    wb, fname, det, pid, params, cambios, registro, contenido, carpeta

donde `params` es el `motor.Parametros` del fichero (§1.2), `contenido` es el
módulo de contenido del producto para ese grupo y `registro` es una lista donde
el grupo puede apuntar `(hoja, celda, formula)` de lo que escriba sin pasar por
`motor.f()`.

CONTENIDO POR PRODUCTO
----------------------
`planes-v2_0/contenido_<pid con guiones bajos>/a.py|b.py|c.py`. Un módulo por
grupo y producto; expone el contenido como **dict del módulo** (`CONTENIDO`, o
las variables en mayúsculas si no lo define). Es lo que traduce el mapa del
representante al molde del hermano: nombres de hoja, filas y cifras del
concepto, nunca lógica.

Térmica: todo en SERIE, un python cada vez. Sin builds, sin navegador.
"""
import argparse
import contextlib
import datetime
import importlib
import inspect
import json
import logging
import os
import re
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    '9a6ebdb7-5d45-4ab1-9e93-10b86cb95c42/scratchpad/planes')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')
SPEC = 'scripts/productos-digitales/planes-v2-SPEC.md §1'

#: Los 10 productos de la familia (censo propio, 2026-08-29).
PRODUCTOS = (
    'plan-negocio-bar-restaurante',
    'plan-negocio-tapas-bar',
    'plan-negocio-cafeteria',
    'plan-negocio-panaderia',
    'plan-negocio-food-truck',
    'plan-negocio-cocteleria-eventos',
    'plan-negocio-parrillero-asador-eventos',
    'plan-negocio-paellero-eventos',
    'plan-catering-tematico-eventos',
    'plan-chef-privado-showcooking-eventos',
)


def _json_seguro(o):
    """Escalares de numpy → tipos nativos; lo demás, a texto."""
    for metodo in ('item', 'tolist'):
        if hasattr(o, metodo):
            try:
                return getattr(o, metodo)()
            except Exception:                                 # noqa: BLE001
                pass
    return str(o)


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Preparación
# ==========================================================================
def preparar_copia(origen, destino):
    if not os.path.isdir(origen):
        raise SystemExit('No existe el origen: ' + origen)
    if os.path.isdir(destino):
        shutil.rmtree(destino)
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    shutil.copytree(origen, destino)
    log('  copia de trabajo regenerada: ' + destino)


def xlsx_de(carpeta):
    return sorted(n for n in os.listdir(carpeta)
                  if n.lower().endswith('.xlsx') and not n.startswith('~$'))


def docx_de(carpeta):
    return sorted(n for n in os.listdir(carpeta)
                  if n.lower().endswith('.docx') and not n.startswith('~$'))


def digest(path):
    """Huella comparable de un .xlsx (valor, formato, relleno, bloqueo,
    merges, DV, formato condicional, protección, impresión)."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked),
                                        str(c.hyperlink.target)
                                        if c.hyperlink else None)
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted(str(dv.type) + ':' + str(dv.formula1) + ':'
                         + str(dv.sqref)
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted(str(cf.sqref) + ':' + str(len(cf.rules))
                         for cf in ws.conditional_formatting),
            'prot': bool(ws.protection.sheet),
            'altos': sorted((str(k), str(v.height))
                            for k, v in ws.row_dimensions.items()
                            if v.height),
            'pie': str(ws.oddFooter.center.text),
        }
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append(fichero + ':' + hoja + ': hoja sólo en una pasada')
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'prot', 'altos', 'pie'):
            if ha[k] != hb[k]:
                fuera.append(fichero + ':' + hoja + ': cambia ' + k + ' ('
                             + str(ha[k])[:140] + ' → ' + str(hb[k])[:140]
                             + ')')
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append(fichero + ':' + hoja + '!' + coord + ': '
                             + str(ca.get(coord)) + ' → ' + str(cb.get(coord)))
    return fuera


# ==========================================================================
# Invocación de los ganchos de los grupos (por nombre de parámetro)
# ==========================================================================
def _llamar(fn, **disponibles):
    """Pasa a `fn` sólo los argumentos que declara.

    Así un grupo puede escribir `def post(wb, fname, cambios)` o
    `def post(wb, fname, det, params, cambios, registro, contenido)` sin que
    main tenga que conocer su firma. Si el grupo usa `**kwargs`, se le pasa
    todo.
    """
    try:
        sig = inspect.signature(fn)
    except (TypeError, ValueError):                          # noqa: BLE001
        return fn(**disponibles)
    if any(p.kind == p.VAR_KEYWORD for p in sig.parameters.values()):
        return fn(**disponibles)
    kwargs = {}
    for nombre, par in sig.parameters.items():
        if par.kind in (par.VAR_POSITIONAL,):
            continue
        if nombre in disponibles:
            kwargs[nombre] = disponibles[nombre]
        elif par.default is par.empty:
            raise TypeError(
                'el gancho ' + getattr(fn, '__name__', '?') + ' pide un '
                'argumento que main no sabe servir: ' + repr(nombre)
                + ' (disponibles: ' + ', '.join(sorted(disponibles)) + ')')
    return fn(**kwargs)


def _ficheros_del_grupo(g, dets, contenido):
    fn = getattr(g, 'ficheros', None)
    if callable(fn):
        return list(_llamar(fn, dets=dets, det_por_fichero=dets,
                            contenido=contenido))
    return list(getattr(g, 'FICHEROS', []))


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, pid, grupos, contenidos, dets, informe_ficheros):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    det = motor.detectar(wb, fname)          # §1.1 — aborta si no lo reconoce
    dets[fname] = det
    # Censo de fórmulas ANTES de tocar nada: el motor puede añadir, nunca
    # perder. Con este contador se cazó que `neutralizar_pseudoformulas`
    # tokenizaba mal `=D6/$D$15` y borraba las 9 fórmulas de porcentaje de
    # `'Resumen'!E6:E14` — 128 → 119 sin que ningún otro gate lo notara.
    antes_formulas = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    antes_formulas[(motor.norm(ws.title),
                                    c.coordinate)] = c.value
    motor.reset(fname)
    cambios, registro_grupo = [], []
    detalle = {}

    activos = [g for g in grupos
               if fname in _ficheros_del_grupo(g, dets, contenidos.get(
                   getattr(g, 'LETRA', '')))]
    propio = any(fname in getattr(g, 'PROPIOS', []) for g in activos)

    base = dict(wb=wb, fname=fname, det=det, pid=pid, cambios=cambios,
                registro=registro_grupo, carpeta=carpeta, params=None,
                contenido=None)

    for g in activos:
        if hasattr(g, 'pre'):
            base['contenido'] = contenidos.get(getattr(g, 'LETRA', ''))
            _llamar(g.pre, **base)

    params = None
    if not propio:
        params, detalle = motor.aplicar(wb, fname, det, pid, cambios)
    base['params'] = params

    for g in activos:
        if hasattr(g, 'post'):
            base['contenido'] = contenidos.get(getattr(g, 'LETRA', ''))
            _llamar(g.post, **base)

    if not propio:
        motor.cerrar(wb, fname, det, pid, cambios, detalle)

    cuenta = motor.contadores(wb, fname, det)

    despues = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    despues.add((motor.norm(ws.title), c.coordinate))
    neutralizadas = set((motor.norm(h), coord)
                        for h, coord, _v, _n in detalle.get(
                            'pseudo_formulas', []))
    perdidas = [{'hoja': h, 'celda': coord, 'formula': antes_formulas[(h,
                                                                      coord)]}
                for (h, coord) in sorted(set(antes_formulas) - despues)
                if (h, coord) not in neutralizadas]
    cuenta['formulas_antes'] = len(antes_formulas)
    cuenta['formulas_perdidas'] = perdidas

    wb.save(path)

    registro = list(motor.REGISTRO)
    registro += [(h, c, fm) for h, c, fm in registro_grupo
                 if isinstance(fm, str) and fm.startswith('=')]
    informe_ficheros.append({
        'fichero': fname, 'tipo': det['tipo'], 'molde': det['molde'],
        'propio_de_grupo': propio,
        'grupos': [getattr(g, 'LETRA', g.__name__) for g in activos],
        'cambios': cambios,
        'formulas_nuevas': len(registro),
        'formatos_corregidos': detalle.get('formatos', []),
        'ortografia_corregida': detalle.get('ortografia', []),
        'totales_descuadrados': detalle.get('desajustes_total', []),
        'cross_sell_sin_precio': detalle.get('cross_sell', []),
        'hojas_sin_inputs_sin_proteger': detalle.get(
            'hojas_sin_inputs_sin_proteger', []),
        'contadores': cuenta,
    })
    return registro


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        salida = r.stdout.decode('utf-8', 'replace').strip()
        fuera.append((n, salida, r.returncode))
        log('    ' + (salida.splitlines()[-1] if salida else '(sin salida)'))
        if r.returncode != 0:
            log('    ERROR inject_cache: '
                + r.stderr.decode('utf-8', 'replace')[-400:])
    return fuera


def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA tiene que tener valor cacheado.

    Una celda vacía sólo se acepta si su fórmula puede devolver `""` **y**
    pycel confirma que ésa es la rama que gana: si devuelve un número y se
    entregó sin caché, el cliente la ve en blanco en Vista previa, en el móvil
    y en Google Sheets, que es el incidente de julio de 2026.
    """
    fallos, vacias, ok = [], 0, 0
    sospechosas, no_verificadas = {}, []
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        wbv = openpyxl.load_workbook(path, data_only=True)
        for hoja, coord, formula in registro:
            ws = None
            for w in wbv.worksheets:
                if motor.norm(w.title) == motor.norm(hoja):
                    ws = w
                    break
            if ws is None:
                fallos.append(fname + ':' + hoja + '!' + coord
                              + ': hoja ausente tras el post-proceso')
                continue
            v = ws[coord].value
            if v is None:
                if '""' in formula:
                    vacias += 1
                    sospechosas.setdefault(fname, []).append((ws.title, coord,
                                                              formula))
                else:
                    fallos.append(fname + ':' + hoja + '!' + coord
                                  + ': sin caché (' + formula[:70] + ')')
            else:
                ok += 1
    comprobadas = 0
    for fname, celdas in sospechosas.items():
        path = os.path.join(carpeta, fname)
        try:
            xl = _pycel(path)
        except Exception as e:                               # noqa: BLE001
            no_verificadas.append(fname + ': no se pudo compilar ('
                                  + type(e).__name__ + ')')
            continue
        for hoja, coord, formula in celdas:
            ref = "'" + hoja + "'!" + coord
            valor = _ev(xl, ref)
            comprobadas += 1
            if isinstance(valor, str) and valor.startswith('ERR:'):
                no_verificadas.append(fname + ':' + ref + ': ' + valor)
            elif valor not in ('', None):
                fallos.append(fname + ':' + ref + ': la fórmula devuelve '
                              + repr(valor) + ' pero se entregó SIN caché ('
                              + formula[:60] + ')')
    return {'con_valor': ok, 'vacias_por_diseno': vacias,
            'vacias_comprobadas_con_pycel': comprobadas,
            'vacias_no_verificadas': no_verificadas, 'fallos': fallos}


# --------------------------------------------------------------------------
# Gate «blancos contaminados» (CRIT-01) — 0 exigido
# --------------------------------------------------------------------------
#: La envoltura estándar de §1.2: `IFERROR(IF(<guarda>="","", CUERPO),"")`.
RX_ENVOLTURA = re.compile(r'^IFERROR\((.*),""\)$', re.S)
RX_GUARDA = re.compile(r'^IF\((.*?)="","",(.*)\)$', re.S)
RX_REF_SIMPLE = re.compile(r"^(?:'([^']+)'!)?\$?([A-Z]{1,3})\$?(\d+)$")


def _guarda_y_cuerpo(formula):
    """Separa la guarda de §1.2 del CUERPO de la fórmula.

    Devuelve `(guarda, cuerpo)`; la guarda es `None` si la fórmula no lleva
    la forma canónica `IF(<celda>="","", …)`.
    """
    s = formula[1:]
    m = RX_ENVOLTURA.match(s)
    if m:
        s = m.group(1)
    m = RX_GUARDA.match(s)
    if m:
        return m.group(1), m.group(2)
    return None, s


def blancos_contaminados(carpeta, nombres):
    """CRIT-01 — celdas en blanco por CONTAMINACIÓN, no por diseño.

    En Excel `número + ""` es `#¡VALOR!`, y el `IFERROR` de la envoltura lo
    convierte en `""`: la celda queda muda y ningún gate se entera. Le pasó a
    la hoja de tesorería del representante —la que el propio libro anuncia
    como «la primera que mira un banco»—, que quedó EN BLANCO de los meses 4
    a 12 porque el flujo del mes sumaba con `+` el pago del IVA, que vale `""`
    en los meses sin liquidación. `verificar_cache` la daba por buena: su
    prueba es `'""' in formula`, y la envoltura SIEMPRE trae un `""`.

    La regla que sí distingue los dos casos mira tres cosas a la vez:

    1. la celda no tiene valor cacheado (`data_only` devuelve `None`);
    2. su CUERPO —quitadas la envoltura y la guarda— no contiene ningún
       literal `""` propio, así que nada en ella estaba pensado para salir en
       blanco;
    3. su guarda es FALSA (la celda de la que cuelga la fila sí tiene valor),
       luego el blanco tampoco viene de ahí.

    Las tres juntas sólo se cumplen cuando un operando contaminó la suma.
    Calibrado sobre el representante: 36 candidatas brutas, de las que 24 son
    los años fuera del plazo del préstamo en `7. Financiación` (guarda cierta)
    y 12 son el defecto real.
    """
    fuera = []
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        wf = openpyxl.load_workbook(path)
        wd = openpyxl.load_workbook(path, data_only=True)
        for ws in wf.worksheets:
            wsd = wd[ws.title]
            for row in ws.iter_rows():
                for c in row:
                    formula = c.value
                    if not motor._es_formula(formula):
                        continue
                    if wsd[c.coordinate].value is not None:
                        continue
                    guarda, cuerpo = _guarda_y_cuerpo(formula)
                    if '""' in cuerpo:
                        continue            # blanco por diseño
                    if guarda is not None:
                        m = RX_REF_SIMPLE.match(guarda.replace('$', ''))
                        if m:
                            hoja = m.group(1) or ws.title
                            coord = m.group(2) + m.group(3)
                            try:
                                v = wd[hoja][coord].value
                            except KeyError:
                                v = 'HOJA-AUSENTE'
                            if v is None or v == '':
                                continue    # la guarda manda: blanco legítimo
                    fuera.append({'fichero': fname, 'hoja': ws.title,
                                  'celda': c.coordinate,
                                  'formula': formula[:160],
                                  'cuerpo': cuerpo[:110]})
    return fuera


def gates_spec9(carpeta, nombres, dets):
    """Los gates de §9 que se miden sobre los xlsx."""
    res = {'no_latinos': [], 'ortografia': [], 'formatos': [],
           'referencias': [], 'sin_formulas': [], 'sin_bio': [],
           'sin_version': [], 'hojas_sin_a4': [], 'checklists': [],
           'sin_instrucciones': [], 'creator_mal': [], 'nombres_hoja': [],
           'cf_desanclado': []}
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        wb = openpyxl.load_workbook(path)
        det = dets.get(fname) or motor.detectar(wb, fname)
        res['no_latinos'] += motor.gate_no_latinos(wb, fname)
        res['ortografia'] += motor.gate_ortografia(wb, fname)
        res['formatos'] += motor.gate_formatos(wb, fname)
        res['referencias'] += motor.gate_referencias(wb, fname)
        res['nombres_hoja'] += motor.gate_nombres_hoja(wb, fname)
        res['cf_desanclado'] += motor.gate_cf_anclado(wb, fname)
        formulas = 0
        bio = version = False
        for ws in wb.worksheets:
            if ws.page_setup.paperSize != 9:
                res['hojas_sin_a4'].append(fname + ':' + ws.title)
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str) and v.startswith('='):
                        formulas += 1
                    elif isinstance(v, str):
                        if motor.RX_BIO.search(v):
                            bio = True
                        if motor.RX_VERSION.match(v) and '2.0' in v:
                            version = True
        if formulas == 0:
            res['sin_formulas'].append(fname)
        if not bio:
            res['sin_bio'].append(fname)
        if not version:
            res['sin_version'].append(fname)
        if motor.hoja(wb, 'Instrucciones') is None:
            res['sin_instrucciones'].append(fname)
        if (wb.properties.creator or '') != motor.CREATOR:
            res['creator_mal'].append(fname + ': '
                                      + repr(wb.properties.creator))
        if det['tipo'] == 'checklist':
            items = 0
            con_dv = 0
            for ws in wb.worksheets:
                for dv in ws.data_validations.dataValidation:
                    if dv.type == 'list' and motor.MARCA_OK in str(dv.formula1):
                        for r in dv.sqref.ranges:
                            cr = r
                            items += ((cr.max_row - cr.min_row + 1)
                                      * (cr.max_col - cr.min_col + 1))
                            con_dv += 1
            res['checklists'].append({'fichero': fname, 'molde': det['molde'],
                                      'items_con_desplegable': items,
                                      'rangos_dv': con_dv})
    return res


def gate_docx(carpeta):
    """Solo lectura: §9 exige A4 + `author='AI Chef Pro'` en los 46 docx.

    Los produce `documentos.py` (§4), que va DESPUÉS de los xlsx: aquí se
    miden y se listan como PENDIENTES, no como fallo del motor. Cantar verde
    sin medirlos sería el mecanismo que dejó pasar la v1.1.
    """
    fuera = []
    for n in docx_de(carpeta):
        fuera.append(motor.censar_docx(os.path.join(carpeta, n)))
    return fuera


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], stdout=subprocess.PIPE,
                       stderr=subprocess.PIPE)
    salida = r.stdout.decode('utf-8', 'replace').strip()
    err = r.stderr.decode('utf-8', 'replace').strip()
    log('  ' + (salida.splitlines()[-1] if salida else '(sin salida)'))
    if r.returncode != 0 and err:
        log('  ' + err[-900:])
    return {'exit': r.returncode, 'salida': salida.splitlines()[-6:],
            'stderr': err.splitlines()[-12:]}


# ==========================================================================
# Demostraciones (pycel) — sobre copias desechables, nunca sobre entregables
# ==========================================================================
def demo_anualidad():
    """Cuota francesa sin `PMT`: 100.000 € al 5 % a 5 años → 1.887,12 €/mes."""
    cuota = motor.cuota_anualidad(100000.0, 0.05, 5)
    return {'cuota_algebraica': round(cuota, 2), 'esperado': 1887.12,
            'ok': abs(round(cuota, 2) - 1887.12) < 0.01,
            'motivo': 'pycel 1.0b30 NO implementa PMT (§ cabecera SPEC)'}


def demo_tir_van_payback():
    """−150.000 / 30.000 / 45.000 / 60.000 / 70.000 → TIR 11,9592 %."""
    flujos = [-150000.0, 30000.0, 45000.0, 60000.0, 70000.0]
    tir = motor.tir_newton(flujos)
    return {'flujos': flujos,
            'tir_pct': round(tir * 100, 4) if tir is not None else None,
            'esperado_pct': 11.9592,
            'van_8pct': round(motor.van(0.08, flujos), 2),
            'payback_anios': motor.payback(flujos),
            'ok': tir is not None and abs(round(tir * 100, 4) - 11.9592) < 0.01,
            'motivo': 'pycel NO implementa IRR: se cachea por Newton'}


def demo_countif_sin_counta(demos_dir):
    """`COUNTIF(r,"<>")` evalúa en pycel; `COUNTA` no existe."""
    os.makedirs(demos_dir, exist_ok=True)
    path = os.path.join(demos_dir, 'demo-countif.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'D'
    for i, v in enumerate(['a', 'b', None, 'd'], start=1):
        ws.cell(row=i, column=1, value=v)
    ws['B1'] = '=COUNTIF(A1:A4,"<>")'
    ws['B2'] = '=COUNTIF(A1:A4,"?*")'
    wb.save(path)
    xl = _pycel(path)
    return {'countif_no_vacio': _ev(xl, "'D'!B1"),
            'countif_texto': _ev(xl, "'D'!B2"),
            'ok': _ev(xl, "'D'!B2") == 3,
            'motivo': 'sustituto obligatorio de COUNTA (§ cabecera SPEC)'}


def demo_iferror(carpeta, nombres, demos_dir):
    """Una división guardada devuelve `""`, no `#¡DIV/0!`, con el input vacío."""
    os.makedirs(demos_dir, exist_ok=True)
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if not (isinstance(v, str) and v.upper().startswith(
                            '=IFERROR(') and '/' in v):
                        continue
                    destino = os.path.join(demos_dir, 'iferror-' + fname)
                    shutil.copy2(path, destino)
                    wb2 = openpyxl.load_workbook(destino)
                    ws2 = wb2[ws.title]
                    # se vacían los precedentes de la división para forzar el
                    # #¡DIV/0! que la guarda tiene que atrapar
                    import re as _re
                    for ref in _re.findall(r'(?<![A-Z0-9_$!])\$?[A-Z]{1,3}\$?'
                                           r'\d+', v.split('/')[-1]):
                        try:
                            ws2[ref.replace('$', '')] = None
                        except Exception:                    # noqa: BLE001
                            pass
                    wb2.save(destino)
                    xl = _pycel(destino)
                    valor = _ev(xl, "'" + ws.title + "'!" + c.coordinate)
                    return {'fichero': fname, 'hoja': ws.title,
                            'celda': c.coordinate, 'formula': v[:90],
                            'valor_con_divisor_vacio': repr(valor),
                            'ok': valor in ('', 0, None)
                            or not str(valor).startswith('#'),
                            'motivo': '§1.5: toda división lleva IFERROR(…,"")'}
    return {'ok': True, 'motivo': 'no hay divisiones en este producto'}


def demo_moldes():
    """Los 30 xlsx de los 10 productos se clasifican (§1.1). Solo LECTURA."""
    fuera, fallos, largos = [], [], []
    for pid in PRODUCTOS:
        carpeta = os.path.join(DL, pid)
        if not os.path.isdir(carpeta):
            fallos.append('no existe ' + carpeta)
            continue
        for n in xlsx_de(carpeta):
            try:
                wb = openpyxl.load_workbook(os.path.join(carpeta, n))
                det = motor.detectar(wb, n)
                fuera.append({'producto': pid, 'fichero': n,
                              'tipo': det['tipo'], 'molde': det['molde']})
                # RC-29 — el aviso de openpyxl («Title is more than 31
                # characters») se pierde entre la salida del run y se queda
                # sin reportar; al reescribir el fichero, Excel puede pedir
                # reparación. Aquí se MIDE sobre los 30 xlsx de la familia.
                largos += motor.gate_nombres_hoja(wb, pid + '/' + n)
            except motor.MoldeDesconocido as e:              # noqa: BLE001
                fallos.append(str(e))
    moldes = {}
    for r in fuera:
        moldes[r['molde']] = moldes.get(r['molde'], 0) + 1
    return {'xlsx_clasificados': len(fuera), 'por_molde': moldes,
            'detalle': fuera, 'fallos': fallos,
            'nombres_de_hoja_demasiado_largos': largos,
            'ok': not fallos and len(fuera) == 30}


def demo_proteccion(carpeta, nombres):
    """Protección sin contraseña y verdes desbloqueadas.

    Una hoja de datos SIN un solo input declarado no se protege (la dejaría
    inservible: el cliente no podría teclear nada hasta que §2/§3 marquen los
    verdes). Eso es un PENDIENTE del grupo, no un fallo del motor; lo que sí
    es fallo es una hoja con inputs sin proteger, una protegida con contraseña
    o una celda verde bloqueada.
    """
    total, protegidas, sin_verdes, fallos = 0, 0, [], []
    pendientes = []
    for fname in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wb.worksheets:
            total += 1
            tiene_input = any(motor.es_verde(c) or not c.protection.locked
                              for row in ws.iter_rows() for c in row
                              if c.value is not None)
            debe = tiene_input or not any(
                isinstance(c.value, (int, float))
                and not isinstance(c.value, bool)
                for row in ws.iter_rows() for c in row)
            if ws.protection.sheet:
                protegidas += 1
            elif debe:
                fallos.append(fname + ':' + ws.title + ': hoja sin proteger')
            else:
                pendientes.append(fname + ':' + ws.title
                                  + ': sin proteger porque aún no tiene '
                                    'ningún input marcado (§2/§3)')
            if ws.protection.password:
                fallos.append(fname + ':' + ws.title
                              + ': protegida CON contraseña (§ convenciones: '
                                'sin contraseña)')
            verdes = sum(1 for row in ws.iter_rows() for c in row
                         if motor.es_verde(c))
            bloqueadas = sum(1 for row in ws.iter_rows() for c in row
                             if motor.es_verde(c) and c.protection.locked)
            if verdes and bloqueadas:
                fallos.append(fname + ':' + ws.title + ': ' + str(bloqueadas)
                              + ' celdas verdes BLOQUEADAS')
            if not verdes:
                sin_verdes.append(fname + ':' + ws.title)
    return {'total_hojas': total, 'protegidas': protegidas,
            'hojas_sin_celdas_verdes': sin_verdes,
            'sin_proteger_por_falta_de_inputs': pendientes, 'fallos': fallos}


def demo_bio_version(carpeta, nombres, pid):
    fallos, ok = [], 0
    for fname in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        bio = version = False
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if not isinstance(c.value, str):
                        continue
                    if motor.RX_BIO.search(c.value):
                        bio = True
                    if motor.RX_VERSION.match(c.value) and '2.0' in c.value \
                            and pid in c.value:
                        version = True
        if bio and version:
            ok += 1
        else:
            fallos.append(fname + ': ' + ('sin bio ' if not bio else '')
                          + ('sin línea de versión 2.0 con el pid'
                             if not version else ''))
    return {'con_bio_y_version': ok, 'total': len(nombres), 'fallos': fallos}


# ==========================================================================
# main
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 de la familia Planes de Negocio')
    ap.add_argument('--producto', default='plan-negocio-bar-restaurante',
                    help='id de carpeta en astro-site/public/dl/')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    ap.add_argument('--solo', default='motor,a,b,c',
                    help='etapas: motor (ninguno), a, b, c o «a,b»')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--origen', default=None,
                    help='carpeta de origen alternativa (exige --dry-run)')
    ap.add_argument('--sin-idempotencia', action='store_true')
    ap.add_argument('--sin-demos', action='store_true')
    args = ap.parse_args()

    pid = args.producto
    if args.origen and not args.dry_run:
        raise SystemExit('ABORTADO: --origen sólo se admite con --dry-run.')
    origen = args.origen or os.path.join(DL, pid)
    destino = os.path.join(SCRATCH, 'dryrun', pid)
    idem_dir = os.path.join(SCRATCH, 'dryrun', pid + '-idem')
    demos_dir = os.path.join(SCRATCH, 'dryrun', '_demos', pid)

    if not args.dry_run and os.environ.get('PLANES_APPLY') != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en ' + origen
            + ', que la SPEC declara intocable hasta que T5/T5\' firmen. Usa '
              '--dry-run (o PLANES_APPLY=1 si eres el orquestador y ya tienes '
              'el visto bueno).')

    respaldo = None
    if args.dry_run:
        preparar_copia(origen, destino)
        carpeta = destino
    else:
        carpeta = origen
        respaldo = os.path.join(
            SCRATCH, 'respaldos', pid + '.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(origen, respaldo)
        log('  respaldo previo de los entregables: ' + respaldo)

    pedidos = [g.strip().lower() for g in args.solo.split(',') if g.strip()]
    solo_motor = all(g in ('motor', 'ninguno', 'none', '') for g in pedidos)
    grupos, cargados, ausentes, contenidos = [], [], [], {}
    paquete_contenido = 'contenido_' + pid.replace('-', '_')
    for letra in pedidos:
        if letra in ('motor', 'ninguno', 'none', ''):
            continue
        try:
            g = importlib.import_module('grupo_' + letra)
        except ImportError as e:                             # noqa: BLE001
            # Un grupo PEDIDO que no carga es AMARILLO mientras no exista
            # (T1 construye sólo el motor) y ROJO en cuanto el paquete diga
            # que está: `main.py` lo distingue por si el fichero existe.
            existe = os.path.isfile(os.path.join(AQUI,
                                                 'grupo_' + letra + '.py'))
            ausentes.append({'grupo': letra, 'fichero_existe': existe,
                             'error': str(e),
                             'nivel': 'fallo' if existe else 'pendiente'})
            log('  grupo_' + letra + (' NO carga: ' + str(e) if existe
                                      else ' aún no existe (pendiente)'))
            continue
        if not hasattr(g, 'LETRA'):
            g.LETRA = letra
        grupos.append(g)
        cargados.append(letra)
        log('  grupo_' + letra + ' cargado')
        try:
            contenidos[letra] = importlib.import_module(
                paquete_contenido + '.' + letra)
            log('  ' + paquete_contenido + '.' + letra + ' cargado')
        except ImportError:
            contenidos[letra] = None
            log('  ' + paquete_contenido + '.' + letra
                + ' no existe — el grupo trabaja sin contenido propio')
        if hasattr(g, 'configurar'):
            _llamar(g.configurar, pid=pid, contenido=contenidos.get(letra),
                    carpeta=carpeta)

    nombres = xlsx_de(carpeta)
    dets = {}
    informe_ficheros, registros = [], {}
    log('\n== 1/8 · post-proceso de ' + str(len(nombres)) + ' xlsx de ' + pid
        + ' ==')
    try:
        for fname in nombres:
            registros[fname] = procesar(carpeta, fname, pid, grupos,
                                        contenidos, dets, informe_ficheros)
            log('  ' + fname + ' [' + dets[fname]['molde'] + ']: '
                + str(len(registros[fname])) + ' fórmulas nuevas')
    except motor.MoldeDesconocido as e:                      # noqa: BLE001
        log('\nABORTADO (§1.1) — ' + str(e))
        if args.json:
            os.makedirs(os.path.dirname(os.path.abspath(args.json)),
                        exist_ok=True)
            with open(args.json, 'w', encoding='utf-8') as fh:
                json.dump({'producto': pid, 'version': '2.0', 'spec': SPEC,
                           'rol': 'motor',
                           'modo': 'dry-run' if args.dry_run else 'produccion',
                           'abortado': str(e), 'fallos': [str(e)], 'exit': 2},
                          fh, ensure_ascii=False, indent=1)
        return 2

    # ---- idempotencia ---------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 2/8 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(idem_dir):
            shutil.rmtree(idem_dir)
        shutil.copytree(carpeta, idem_dir)
        antes = dict((n, digest(os.path.join(carpeta, n))) for n in nombres)
        for fname in nombres:
            procesar(idem_dir, fname, pid, grupos, contenidos, {}, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(idem_dir, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs), 'detalle': difs}
        log('  diferencias 1.ª vs 2.ª pasada: ' + str(len(difs)))
        for d in difs[:10]:
            log('    ' + d)

    # ---- caché ----------------------------------------------------------
    log('\n== 3/8 · inject_cache (siempre el último que toca el zip) ==')
    cache = inject_cache(carpeta, nombres)

    log('\n== 4/8 · verificación data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log('  con valor: ' + str(ver['con_valor']) + ' · "" por diseño: '
        + str(ver['vacias_por_diseno']) + ' · fallos: '
        + str(len(ver['fallos'])))
    for fl in ver['fallos'][:10]:
        log('    ' + fl)
    # CRIT-01 — el gate que `verificar_cache` no puede ser: aquí no vale
    # preguntar si la fórmula «puede» devolver "", porque la envoltura de
    # §1.2 siempre trae uno. 0 exigido.
    blancos = blancos_contaminados(carpeta, nombres)
    log('  blancos_contaminados: ' + str(len(blancos)))
    for b in blancos[:10]:
        log('    ' + b['fichero'] + ':' + b['hoja'] + '!' + b['celda']
            + ' → ' + b['cuerpo'])

    log('\n== 5/8 · gates de §9 sobre los xlsx ==')
    g9 = gates_spec9(carpeta, nombres, dets)
    G9_CLAVES = ('no_latinos', 'ortografia', 'formatos', 'referencias',
                 'sin_formulas', 'sin_bio', 'sin_version',
                 'sin_instrucciones', 'hojas_sin_a4', 'creator_mal',
                 'nombres_hoja', 'cf_desanclado')
    for clave in G9_CLAVES:
        log('  ' + clave + ': ' + str(len(g9[clave])))
    limpios = (sum(1 for k in G9_CLAVES if not g9[k])
               + (0 if blancos else 1))
    log('  --- gates medibles a 0: ' + str(limpios) + '/'
        + str(len(G9_CLAVES) + 1) + ' (los 12 de §9 + blancos_contaminados)')
    for c in g9['checklists']:
        log('  checklist ' + c['fichero'] + ' [' + c['molde'] + ']: '
            + str(c['items_con_desplegable']) + ' ítems con desplegable')

    log('\n== 6/8 · docx (solo lectura: los produce documentos.py, §4) ==')
    docx = gate_docx(carpeta)
    pend_docx = [d for d in docx if not d.get('a4')
                 or d.get('author') != motor.CREATOR]
    log('  ' + str(len(docx)) + ' docx · pendientes de A4/autor: '
        + str(len(pend_docx)))

    log('\n== 7/8 · censo-entregables --fail ==')
    cen = censo(carpeta)

    demos = {}
    if not args.sin_demos:
        log('\n== 8/8 · demostraciones (pycel) ==')
        demos = {
            'anualidad_sin_PMT': demo_anualidad(),
            'tir_van_payback_sin_IRR': demo_tir_van_payback(),
            'countif_sin_COUNTA': demo_countif_sin_counta(demos_dir),
            'iferror_vivo': demo_iferror(carpeta, nombres, demos_dir),
            'deteccion_de_moldes': demo_moldes(),
            'proteccion': demo_proteccion(carpeta, nombres),
            'bio_y_version': demo_bio_version(carpeta, nombres, pid),
        }
        log('  cuota algebraica 100.000/5 %/5 años: '
            + str(demos['anualidad_sin_PMT']['cuota_algebraica']) + ' €')
        log('  TIR del caso trazado: '
            + str(demos['tir_van_payback_sin_IRR']['tir_pct']) + ' %')
        log('  moldes clasificados: '
            + str(demos['deteccion_de_moldes']['xlsx_clasificados']) + '/30 '
            + str(demos['deteccion_de_moldes']['por_molde']))
        log('  hojas protegidas: ' + str(demos['proteccion']['protegidas'])
            + '/' + str(demos['proteccion']['total_hojas']))

    # ---- veredicto ------------------------------------------------------
    fallos = []
    for g in grupos:
        if hasattr(g, 'demos'):
            propias = _llamar(g.demos, carpeta=carpeta, origen=origen,
                              demos_dir=demos_dir, pid=pid)
            if isinstance(propias, dict):
                fallos += propias.pop('fallos', [])
                demos.update(propias)
    for a in ausentes:
        if a['nivel'] == 'fallo':
            fallos.append('grupo_' + a['grupo'] + ' se pidió, su fichero '
                          'existe y NO se pudo cargar (' + a['error']
                          + '): las secciones que le tocan quedan SIN aplicar')
    if idem.get('diferencias'):
        fallos.append('idempotencia: ' + str(idem['diferencias'])
                      + ' diferencias entre la 1.ª y la 2.ª pasada')
    if ver['fallos']:
        fallos.append('caché: ' + str(len(ver['fallos']))
                      + ' fórmulas nuevas sin valor')
    if blancos:
        fallos.append(
            'CRIT-01 / blancos_contaminados: ' + str(len(blancos))
            + ' celdas quedan EN BLANCO por contaminación, no por diseño '
            '(un operando vale "" y en Excel número + "" es #¡VALOR!, que el '
            'IFERROR convierte en blanco): '
            + ', '.join(b['hoja'] + '!' + b['celda'] for b in blancos[:8]))
    for nombre, salida, rc in cache:
        if rc != 0:
            fallos.append('inject_cache falló en ' + nombre + ' (exit '
                          + str(rc) + ')')
        if 'fallos_pycel' in (salida or '') and 'fallos_pycel=0' not in salida:
            fallos.append('inject_cache: ' + nombre + ' → ' + salida[-90:])
    for fi in informe_ficheros:
        perdidas = fi['contadores'].get('formulas_perdidas') or []
        if perdidas:
            fallos.append(fi['fichero'] + ': el post-proceso PERDIÓ '
                          + str(len(perdidas)) + ' fórmulas que el fichero ya '
                          'tenía (' + ', '.join(p['hoja'] + '!' + p['celda']
                                                for p in perdidas[:6]) + ')')
    if g9['no_latinos']:
        fallos.append('§9: ' + str(len(g9['no_latinos']))
                      + ' celdas con caracteres no latinos')
    if g9['referencias']:
        fallos.append('§1.7: ' + str(len(g9['referencias'])) + ' fórmulas '
                      'apuntan a una hoja que no existe tras el renombrado')
    if g9['creator_mal']:
        fallos.append('§1.9: creator distinto de «AI Chef Pro» en '
                      + str(len(g9['creator_mal'])) + ' ficheros')
    # RT-01 — un formato condicional de tipo `expression` cuya fórmula no
    # arranca en la primera fila de su `sqref` pinta la tabla desplazada. Es
    # BLOQUEANTE: en la v1.1 estaba bien y el motor lo rompió.
    if g9['cf_desanclado']:
        fallos.append('RT-01: ' + str(len(g9['cf_desanclado'])) + ' formatos '
                      'condicionales con la fórmula desanclada del sqref ('
                      + ', '.join(x['hoja'] + ' ' + x['celda'] + ' → '
                                  + x['formula']
                                  for x in g9['cf_desanclado'][:4]) + ')')
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió ' + str(cen['exit']))
    if demos:
        for clave in ('anualidad_sin_PMT', 'tir_van_payback_sin_IRR',
                      'countif_sin_COUNTA', 'iferror_vivo',
                      'deteccion_de_moldes'):
            d = demos.get(clave) or {}
            if d and not d.get('ok'):
                fallos.append('demo ' + clave + ' NO pasa: ' + repr(d)[:220])
        fallos += demos.get('proteccion', {}).get('fallos', [])
        fallos += demos.get('bio_y_version', {}).get('fallos', [])

    # Pendientes: lo que este rol NO cierra pero §9 exige antes del LIVE.
    pendientes = []
    if solo_motor:
        pendientes.append('§2/§3: sin grupos cargados, el modelo financiero '
                          'no está construido (0 fórmulas en los libros que '
                          'sólo tenían constantes es esperable aquí)')
    for a in ausentes:
        if a['nivel'] == 'pendiente':
            pendientes.append('grupo_' + a['grupo'] + ' aún no existe')
    if g9['sin_formulas']:
        pendientes.append('§9 «0 xlsx con 0 fórmulas»: siguen sin fórmula '
                          + str(len(g9['sin_formulas'])) + ' ('
                          + ', '.join(g9['sin_formulas']) + ')')
    if g9['formatos']:
        pendientes.append('§1.4: quedan ' + str(len(g9['formatos']))
                          + ' celdas con formato incoherente con su rótulo')
    if g9['ortografia']:
        pendientes.append('§1.7: quedan ' + str(len(g9['ortografia']))
                          + ' textos sin tilde ('
                          + ', '.join(sorted(set(
                              x.get('palabra', '') for x in g9['ortografia']
                          ))[:12]) + ')')
    # RC-29 — los nombres de pestaña de más de 31 caracteres los corrige el
    # módulo de contenido de su producto (parrillero, T8): aquí se MIDEN para
    # que el aviso de openpyxl deje de perderse entre la salida.
    largos = (g9['nombres_hoja']
              + ((demos.get('deteccion_de_moldes') or {})
                 .get('nombres_de_hoja_demasiado_largos') or []))
    if largos:
        pendientes.append('RC-29: ' + str(len(largos)) + ' nombres de hoja de '
                          'más de 31 caracteres, que Excel corta y puede '
                          'obligar a reparar el fichero ('
                          + ', '.join(x['fichero'] + '!' + x['hoja'] + ' ('
                                      + str(x['longitud']) + ')'
                                      for x in largos)
                          + '); el renombrado va en el módulo de contenido de '
                            'su producto')
    sin_prot = (demos.get('proteccion') or {}).get(
        'sin_proteger_por_falta_de_inputs') or []
    if sin_prot:
        pendientes.append('§1.3: ' + str(len(sin_prot)) + ' hojas de datos '
                          'quedan SIN proteger porque ningún grupo les ha '
                          'marcado todavía una celda de input verde '
                          '(protegerlas ahora las dejaría inservibles)')
    if pend_docx:
        pendientes.append('§9: ' + str(len(pend_docx)) + ' docx sin A4 o sin '
                          'author=AI Chef Pro (los produce documentos.py, T9)')

    informe = {
        'producto': pid,
        'version': '2.0',
        'rol': 'motor',
        'spec': SPEC,
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'carpeta_origen': origen,
        'carpeta_trabajo': carpeta,
        'grupos_pedidos': pedidos,
        'grupos_cargados': cargados,
        'grupos_ausentes': ausentes,
        'contenido_cargado': dict((k, bool(v)) for k, v in contenidos.items()),
        'solo_motor': solo_motor,
        'moldes': dict((n, dets[n]['tipo'] + '/' + dets[n]['molde'])
                       for n in dets),
        'ficheros': informe_ficheros,
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'data_only_formulas_nuevas': ver,
            'blancos_contaminados': blancos,
            'spec9': g9,
            'docx_solo_lectura': docx,
            'censo_entregables': cen,
        },
        'demostraciones': demos,
        'fallos': fallos,
        'pendientes': pendientes,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            # pycel devuelve escalares de numpy (`bool_`, `float64`) en
            # cuanto una fórmula pasa por SUMPRODUCT o MATCH, y el informe
            # los arrastra: sin este conversor, el volcado del JSON revienta
            # DESPUÉS de haber hecho todo el trabajo.
            json.dump(informe, fh, ensure_ascii=False, indent=1,
                      default=_json_seguro)
        log('\ninforme → ' + args.json)

    log('\n' + ('FALLOS:\n  ' + '\n  '.join(str(x) for x in fallos) if fallos
                else 'TODO VERDE (moldes, idempotencia, caché, blancos '
                     'contaminados, gates §9, censo, demos)'))
    if pendientes:
        log('\nPENDIENTE (no lo cierra el motor):\n  '
            + '\n  '.join(pendientes))
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (bórralo sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
