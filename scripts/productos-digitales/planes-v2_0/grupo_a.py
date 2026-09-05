#!/usr/bin/env python3
"""
grupo_a.py — Línea A de la familia «Planes de Negocio» v2.0 (§2 de la SPEC).

Convierte el plan financiero de los CINCO productos de línea A en un MODELO:
una sola hoja de entrada (`0. Supuestos`) y todo lo demás derivado por fórmula,
más dos hojas nuevas (`Tesorería 12 meses` y `Financiación`) y el checklist de
apertura con el contenido legal vigente.

    SPEC: scripts/productos-digitales/planes-v2-SPEC.md §2 (§2.1 a §2.12)

POR MOLDE, NUNCA POR POSICIÓN FIJA (§1.1 / §7-bis.9)
----------------------------------------------------
La línea A tiene DOS moldes y este módulo se escribió mirando los tres ficheros
a la vez, no sólo el del representante:

  * **A-α** — `plan-negocio-bar-restaurante`. Hojas numeradas
    (`'1. Inversion Inicial'`…`'5. Personal'`), ingresos = cubiertos × ticket ×
    días en filas de input, hoja `Personal` CON columna «Personas».
  * **A-β** — `tapas-bar`, `cafeteria`, `panaderia`, `food-truck`. Hojas sin
    numerar (`'Inversion Inicial'`, `'PyG 3 Anos'`, …), ingresos = **4 líneas
    tecleadas por familia de producto** y **ningún input de ticket ni de
    cubiertos en el P&L** (viven en `Punto Equilibrio` y en `Escenarios`, con
    dos calendarios distintos — `NUEVO-03`), hoja `Personal` **SIN** columna
    «Personas» y con las notas en la columna F.

Todo acceso a hoja va por `motor.hoja()` (insensible a tildes: §1.7 renombra
`'PyG 3 Anos'` → `'PyG 3 Años'` DESPUÉS de este grupo) y toda fila se localiza
por su RÓTULO normalizado, nunca por su número. Las dos únicas cosas que este
módulo da por sabidas son las que la SPEC fija: la rejilla de `0. Supuestos`
(§2.1) y que la columna A lleva el concepto.

DE DÓNDE SALE CADA NÚMERO
-------------------------
1. `0. Supuestos` — lo escribe este grupo con los valores de
   `contenido_<pid>/a.py`, que es donde vive TODO lo específico del producto
   (importes, rótulos, plantilla, umbrales) con la fuente de cada cifra.
2. El resto del libro se deriva de ahí por fórmula. Ningún número se teclea dos
   veces y ningún literal sobrevive dentro de una fórmula ni de un rótulo
   (§7-bis.11): los rótulos que hoy llevan el porcentaje escrito pasan a
   llevarlo en la NOTA, generada con `TEXT()` desde la celda del parámetro.
3. Las partidas que el fichero ya traía y que son datos del cliente (las líneas
   de inversión, los costes fijos que no son driver) se CONSERVAN como celdas
   de input verdes (§1.3): no se borra ningún número que el cliente pueda estar
   usando. Lo que cambia de valor se anota en el informe (`RECALIBRADO`).

DECISIONES QUE ESTE MÓDULO TOMA Y POR QUÉ
-----------------------------------------
a) **Reconstrucción de hoja, no inserción de filas.** `ws.insert_rows()` de
   openpyxl no mueve validaciones, formato condicional ni celdas combinadas, y
   no reescribe fórmulas: insertar dejaría el libro incoherente. Cada hoja se
   vacía en su zona de datos y se reescribe entera desde el modelo. Es lo que
   hace que la 2.ª pasada sea idéntica a la 1.ª (idempotencia = 0 diferencias,
   que `main.py` comprueba por huella).
b) **Los rótulos NO llevan fórmula; la nota SÍ.** §7-bis.11 exige que los
   números de los rótulos («Food cost (30% sobre ingresos)», «Colchon operativo
   (3 meses…)») dejen de estar escritos a mano. Se cumple quitándolos del
   rótulo y generando la frase con `TEXT()` en la columna de NOTAS. No se
   ponen fórmulas en la columna A porque `motor._rotulo_de_fila()` ignora las
   celdas que empiezan por «=»: el motor perdería el rótulo y con él la regla
   de formatos por tipo (§1.4) y la validación por rótulo (§1.5).
c) **Sin referencia circular.** El fondo de maniobra depende de los costes
   fijos, los costes fijos de los intereses y los intereses del préstamo: si el
   préstamo se derivara de la inversión, Excel daría referencia circular. Por
   eso **recursos propios y préstamo son INPUT** y la hoja `Financiación`
   compara orígenes contra usos con un semáforo (§2.8: «cuadrando con
   `'1. Inversion Inicial'!B46`»).
d) **Cuadro de amortización ANUAL.** La cuota va como anualidad algebraica
   (pycel no implementa `PMT`, §7-bis.16). Se usa periodicidad anual en el
   cuadro y en tesorería se reparte en doce partes iguales, con la nota puesta:
   mezclar un cuadro mensual con un P&L anual descuadra los intereses.
e) **Todo cálculo va envuelto en `IFERROR(...;"")`.** Con el libro en blanco no
   puede quedar ni un semáforo verde ni un «0,0 %» falso: «sin dato» se escribe
   `""` (convención de familia), y el formato condicional lleva la guarda
   `ISNUMBER` que pone el motor (§1.6).
f) **Umbrales al lado del ratio.** El formato condicional compara contra una
   celda de la MISMA hoja (columna «Umbral»), no contra `Instrucciones`: las
   referencias entre hojas dentro de una regla de formato condicional no las
   admiten todas las versiones de Excel, y así además el lector ve el umbral al
   lado del ratio que audita (TEC-12, §2.9).

CONTRATO CON `contenido_<pid>/a.py`
-----------------------------------
Todo es opcional: sin módulo de contenido el grupo sigue funcionando leyendo el
fichero (peor calibrado, pero coherente). Claves que lee:

    CONCEPTO           str — nombre del negocio para los rótulos
    SUPUESTOS          {clave: (coord, etiqueta, valor, fmt, nota, fuente)}
    LINEAS_INGRESO     [(rótulo, peso, 'comida'|'bebida', nota, fuente)]
    PLANTILLA          [(puesto, personas, bruto_mes_total, nota, fuente)]
    FIJOS              {rótulo_norm: (accion, nota)}  accion: 'suprimir'|número
    FIJOS_EXTRA        [(rótulo, importe, nota, fuente)]
    INVERSION          {rótulo_norm: (accion, nota)}
    INVERSION_EXTRA    [(bloque, rótulo, importe, nota, fuente)]
    AMORTIZABLE        {'obra': [regex], 'maquinaria': [regex], 'no': [regex]}
    UMBRALES           [(clave, rótulo, valor, comentario)]
    ESCENARIOS         {'pesimista': (cub, ticket, días), 'optimista': (...)}
    ESTACIONALIDAD     [12 pesos que suman 1]
    CHECKLIST          {'reemplazos', 'altas', 'suprimir', 'fases'}
    INSTRUCCIONES      {'uso': [...], 'referencias': [(rótulo, valor, nota)]}
    RECALIBRADO        [(concepto, valor_v1, valor_v2, motivo)]

IDS DEL R1 QUE CIERRA (mapa §8 de la SPEC)
------------------------------------------
TEC-01/04/05/06/07/10/11/12/16/17/18/19/20/21/22/23/25 · DOM-01/04/05/07/08/09/
10/11/12/13/14/15/17/19/23/24/25/26/30/32/33/34 · COM-03/04/06/08/11/12/13/14/
17/20/21/22/25 · NUEVO-01/02/03/06. Los de §1 (formatos, tildes, altos, DV,
metadata) los cierra `motor.py`; los de §4 y §5, `documentos.py` y T10.
"""
import copy
import math
import os
import re
import shutil

from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

import motor

LETRA = 'a'
SPEC = 'planes-v2-SPEC.md §2'

#: Ficheros que este grupo construye ENTEROS (ninguno: el §1 transversal del
#: motor se aplica a los dos ficheros que toca).
PROPIOS = []

FINO = motor.FINO          # U+202F — SIEMPRE por escape (regla dura 5)
GUION = motor.GUION        # U+2011

# ==========================================================================
# Nombres de hoja por molde (§1.1). A-α numera; A-β no.
# ==========================================================================
HOJAS = {
    'inversion': ('1. Inversion Inicial', 'Inversion Inicial'),
    'pyg': ('2. P&L 3 Anos', 'PyG 3 Anos'),
    'equilibrio': ('3. Punto Equilibrio', 'Punto Equilibrio'),
    'escenarios': ('4. Escenarios', 'Escenarios'),
    'personal': ('5. Personal', 'Personal'),
}
#: Hojas nuevas de §2.7 y §2.8. El prefijo numérico se decide por el molde.
NUEVAS = (('tesoreria', 'Tesorería 12 meses', '6. '),
          ('financiacion', 'Financiación', '7. '))

MOLDES = ('A-alfa', 'A-beta')

# ==========================================================================
# Rejilla de `0. Supuestos` (§2.1). El motor reserva B20:B22 y B37:B40.
# ==========================================================================
#: (clave, coord, etiqueta, valor por defecto, formato, nota).
#: El valor por defecto sólo se usa si el módulo de contenido no trae el suyo y
#: no se puede leer del fichero: es el último recurso, nunca la fuente.
SUPUESTOS_BASE = (
    ('cubiertos_dia', 'B4', 'Cubiertos/día (media del año 1)', None,
     motor.FMT_ENT,
     'Comensales servidos al día de media, contando todos los servicios'),
    ('ticket_medio', 'B5', 'Ticket medio SIN IVA (€)', None, motor.FMT_EUR,
     'Gasto medio por comensal sin IVA. En la nota de al lado tienes el PVP '
     'equivalente con IVA'),
    ('dias_apertura', 'B6', 'Días de apertura al año', None, motor.FMT_ENT,
     'El MISMO dato lo usan el P&L, el punto de equilibrio y los escenarios'),
    ('crec_a2', 'B7', 'Crecimiento de volumen del año 2', 0.10,
     motor.FMT_PCT, 'Se aplica a los cubiertos y al coste de personal'),
    ('crec_a3', 'B8', 'Crecimiento de volumen del año 3', 0.06,
     motor.FMT_PCT, 'Se aplica sobre el año 2'),
    ('pct_comida', 'B11', 'Ventas de COMIDA sobre el total', None,
     motor.FMT_PCT, 'Se calcula sumando el peso de las líneas de comida'),
    ('pct_bebida', 'B12', 'Ventas de BEBIDA sobre el total', None,
     motor.FMT_PCT, 'El resto de las ventas'),
    ('coste_comida', 'B13', 'Coste de mercancía sobre las ventas de COMIDA',
     0.30, motor.FMT_PCT,
     'Food cost real: se aplica SOLO a la comida, nunca al total'),
    ('coste_bebida', 'B14', 'Coste de mercancía sobre las ventas de BEBIDA',
     0.22, motor.FMT_PCT, 'Se aplica SOLO a la bebida'),
    ('pct_consumibles', 'B15', 'Consumibles sobre ventas', 0.015,
     motor.FMT_PCT, 'Servilletas, limpieza, envases, papel de TPV'),
    ('pct_delivery', 'B16', 'Ventas por delivery sobre el total', 0.0,
     motor.FMT_PCT,
     'A CERO por defecto: si no repartes, no arrastras un coste inventado'),
    ('comision_delivery', 'B17', 'Comisión de la plataforma de delivery',
     0.28, motor.FMT_PCT,
     'Se aplica solo sobre las ventas del canal, no sobre el total'),
    ('comision_tpv', 'B18', 'Comisión de los medios de pago', 0.008,
     motor.FMT_PCT, 'Tarjeta y bizum sobre el total facturado'),
    ('alquiler_mes', 'B24', 'Alquiler mensual del local (€)', None,
     motor.FMT_EUR0, 'De aquí salen la fianza, el primer mes y el ratio '
     'alquiler/ventas'),
    ('fianza_meses', 'B25', 'Fianza del alquiler (meses)', 3, motor.FMT_ENT,
     'Meses de renta que pide el arrendador como fianza'),
    ('suministros_mes', 'B26', 'Suministros mensuales de luz, agua y gas (€)', None, motor.FMT_EUR0,
     'Luz, agua y gas'),
    ('seguros_ano', 'B27', 'Seguros (€/año)', None, motor.FMT_EUR0,
     'Responsabilidad civil + multirriesgo del local'),
    ('pct_varios', 'B28', 'Varios e imprevistos sobre ventas', 0.02,
     motor.FMT_PCT, 'Colchón de gasto corriente no presupuestado'),
    ('recursos_propios', 'B30', 'Recursos propios aportados (€)', None,
     motor.FMT_EUR0, 'Capital y aportaciones de los socios'),
    ('prestamo', 'B31', 'Préstamo bancario solicitado (€)', None,
     motor.FMT_EUR0, 'Importe del principal. La hoja de Financiación monta el '
     'cuadro de amortización'),
    ('tipo_prestamo', 'B32', 'Tipo de interés nominal anual', 0.06,
     motor.FMT_PCT, 'Pide oferta a dos entidades antes de fijarlo'),
    ('plazo_prestamo', 'B33', 'Plazo del préstamo (años)', 7, motor.FMT_ENT,
     'Años totales, carencia incluida'),
    ('carencia_prestamo', 'B34', 'Carencia de principal (años)', 1,
     motor.FMT_ENT,
     'Durante la carencia sólo se pagan intereses. Si iguala o supera al '
     'plazo, la hoja la anula'),
    ('meses_fondo', 'B35', 'Fondo de maniobra (meses de costes fijos)', 3,
     motor.FMT_ENT,
     'Las Instrucciones de este libro exigen 3 meses como mínimo'),
    ('iva_soportado', 'B41', 'IVA soportado en compras e inversión', 0.21,
     motor.FMT_PCT, 'Recuperable vía modelo 303, pero hay que adelantarlo'),
    ('bin_inicial', 'B42', 'Bases negativas de ejercicios anteriores (€)', 0,
     motor.FMT_EUR0,
     'Pérdidas pendientes de compensar al empezar (art. 26 LIS)'),
    ('vida_obra', 'B44', 'Vida útil de obra e instalaciones (años)', 10,
     motor.FMT_ENT,
     'Coeficientes de la tabla del art. 12.1 LIS; confírmalo con tu asesor'),
    ('vida_maquinaria', 'B45',
     'Vida útil de maquinaria y mobiliario (años)', 8, motor.FMT_ENT,
     'Coeficientes de la tabla del art. 12.1 LIS; confírmalo con tu asesor'),
    ('ipc', 'B48', 'Subida anual de los costes fijos', 0.0, motor.FMT_PCT,
     'A CERO: las tres columnas están en euros del año 1 (términos reales). '
     'Súbela si quieres proyectar en euros corrientes'),
    # ---- ARRANQUE Y AFORO (RD-17/RC-03, RD-24, RT-14, RD-02) -----------
    ('aforo', 'B51', 'Aforo del local (plazas sentadas + barra)', 56,
     motor.FMT_ENT,
     'De aquí sale la rotación implícita: cubiertos/día ÷ aforo. Cuéntalo '
     'sobre el plano de tu local, no lo copies'),
    ('rampa_mes1', 'B52', 'Actividad del mes 1 sobre la de crucero', 0.55,
     motor.FMT_PCT,
     'Un local que abre no factura desde el primer día lo mismo que a los '
     'seis meses. Con 100 % se elimina la rampa y el año 1 se proyecta a '
     'velocidad de crucero'),
    ('meses_rampa', 'B53', 'Meses hasta alcanzar el régimen de crucero', 6,
     motor.FMT_ENT,
     'La rampa sube en línea recta desde el porcentaje de arriba hasta el '
     '100 % en este número de meses'),
    ('meses_renta_previa', 'B54',
     'Meses de alquiler ANTES de abrir (obra y licencias)', 1,
     motor.FMT_ENT,
     'El local se paga desde que se firma. Estos meses son inversión: NO se '
     'solapan con el P&L, que arranca el día de la apertura'),
    # M20 / REF22-BAR-05 (2026-09-05) — desde A8 el porcentaje se aplica sólo
    # a las partidas de OBRA del bloque, no al bloque entero, y el rótulo del
    # input seguía prometiendo lo contrario: quien leyera sólo esta hoja
    # calcularía 10 % × 53.650 € y no entendería los 3.650 €.
    ('pct_imprevistos', 'B55',
     'Imprevistos sobre las partidas de OBRA del bloque de local (%)',
     0.10, motor.FMT_PCT,
     'Colchón sobre la partida de obra y acondicionamiento. Un banco no '
     'financia una reforma sin él'),
    # ---- COBROS, PAGOS Y DESGLOSE DE IVA (RD-35, RD-12, RD-18) ---------
    ('dias_cobro', 'B58', 'Días medios de cobro', 0, motor.FMT_ENT,
     'En barra y sala se cobra al contado (0 días). Súbelo si trabajas con '
     'plataformas de reserva o de reparto, que liquidan a D+1/D+30'),
    ('dias_pago', 'B59', 'Días medios de pago a proveedor', 30,
     motor.FMT_ENT,
     'Lo que te financian tus proveedores. Es lo primero que se pierde si '
     'te retrasas en un pago'),
    ('mes_paga_1', 'B60', 'Mes de la paga extra de verano', 7, motor.FMT_ENT,
     'Mes del año (1-12) en el que sale de caja la primera paga extra'),
    ('mes_paga_2', 'B61', 'Mes de la paga extra de Navidad', 12,
     motor.FMT_ENT,
     'Mes del año (1-12) en el que sale de caja la segunda paga extra'),
    ('pct_bebida_alc', 'B62',
     'Bebida ALCOHÓLICA sobre el total de bebida', 0.60, motor.FMT_PCT,
     'Peso del alcohol dentro de la línea de bebida. Se usa igual en ventas '
     'y en compras, y es una aproximación (el alcohol suele pesar algo más '
     'en el coste que en la venta). En sala todo se repercute al 10 %, '
     'refrescos incluidos. En la factura del proveedor, el alcohol va al '
     'tipo general; el agua, el café y los zumos naturales al 10 %; los '
     'refrescos y zumos con azúcares o edulcorantes añadidos al 21 % desde '
     '2021, que el libro aproxima al 10 %'),
    # RD-17 (2026-09-05, decisión del dueño heredada de la familia de guías):
    # la bebida alcohólica servida EN SALA tributa al 10 %, no al 21 %. El
    # art. 91.Uno.2.2 de la Ley del IVA grava al tipo reducido los servicios
    # de hostelería y «el suministro de comidas y bebidas para consumir en el
    # acto» sin excluir el alcohol; el 21 % es el tipo general y sólo aplica
    # a la venta para llevar. Antes el blend de bebida leía `iva_general`
    # (B40) y el libro repercutía de más. Parámetro propio, celda verde con
    # nota, en el bloque de desglose del IVA junto al peso que lo usa.
    ('iva_bebida', 'B63',
     'IVA de la bebida ALCOHÓLICA servida en sala', 0.10, motor.FMT_PCT,
     '10 %, igual que el resto del consumo en sala: el art. 91.Uno.2.2.º de '
     'la Ley 37/1992 del IVA grava al tipo reducido los servicios de '
     'hostelería y «el suministro de comidas y bebidas para consumir en el '
     'acto», sin excluir el alcohol. El 21 % es el tipo GENERAL y, fuera '
     'del consumo en el acto, alcanza al alcohol y también a los refrescos y '
     'zumos con azúcares o edulcorantes añadidos vendidos para llevar o a '
     'domicilio (art. 91.Uno.1.1.º de la misma ley). El libro lo aplica '
     'sólo a la parte ALCOHÓLICA de la bebida que sale por el canal de la '
     'celda «Ventas por delivery sobre el total»; los refrescos azucarados '
     'repartidos irían también al general y el libro no los separa'),
    # ---- UMBRALES Y SUELOS DE CONTROL (RD-22, RD-25) -------------------
    ('dscr_min', 'B65', 'DSCR mínimo aceptable', 1.0, motor.FMT_DEC2,
     'Por debajo de 1 el negocio no genera para pagar el préstamo. Ponlo en '
     'el covenant que te exija tu banco'),
    ('dscr_obj', 'B66', 'DSCR objetivo (verde)', 1.25, motor.FMT_DEC2,
     'Lo que suele pedir una entidad para dar el préstamo sin garantías '
     'adicionales'),
    ('holgura_min', 'B67',
     'Holgura mínima sobre el punto de equilibrio', 0.15, motor.FMT_PCT,
     'Cuánto puedes caer sobre lo previsto antes de entrar en pérdidas'),
    ('salario_convenio', 'B68',
     'Suelo salarial anual del convenio provincial (€)', 0, motor.FMT_EUR0,
     'El convenio PROVINCIAL de hostelería, no el SMI, es el suelo real. '
     'Cópialo de la tabla salarial de tu provincia; con 0 el semáforo '
     'compara sólo contra el SMI'),
)

#: Bloques extra que este grupo cuelga debajo de la rejilla del motor.
BLOQUES_EXTRA = (
    ('A47', 'CRECIMIENTO Y ACTUALIZACIÓN DE COSTES'),
    # REF-CIF-06 (2026-09-05): iba en A50 y lo pisaba «Rotaciones al día
    # implícitas (calculado)», así que el bloque salía sin título.
    ('A49', 'ARRANQUE Y AFORO'),
    ('A57', 'COBROS, PAGOS Y DESGLOSE DEL IVA'),
    ('A64', 'UMBRALES Y SUELOS DE CONTROL'),
)
BLOQUE_EXTRA = BLOQUES_EXTRA[0]     # compatibilidad

# ==========================================================================
# Clasificación de partidas (por RÓTULO: sirve para A-α con bloques y para
# A-β, que es una lista plana sin bloques)
# ==========================================================================
RX_CANON_INV = re.compile(
    r'fianza|primer mes.*alquiler|fondo de maniobra|colch[oó]n operativo|'
    r'iva soportado|base amortizable|amortizaci[oó]n anual|'
    r'necesidad total de caja', re.I)
#: ⚠️ «comisiones» a secas se tragaba «Comisiones de reservas online», que es
#: una partida NUEVA y preservable: al excluirla del barrido, la 2.ª pasada la
#: volvía a añadir al final y las dos últimas filas de costes fijos cambiaban
#: de orden. El patrón nombra sólo las dos comisiones que genera este grupo.
RX_CANON_PYG = re.compile(
    r'^(alquiler|n[oó]minas|salarios|personal|suministros|seguros?|'
    r'amortizaci[oó]n|cuota pr[eé]stamo|gastos financieros|'
    r'varios e imprevistos|coste de mercanc[ií]a|food cost|bebidas? cost|'
    r'consumibles|comisiones de (delivery|los medios|medios)|'
    r'ingresos|cubiertos|clientes|ticket|d[ií]as|'
    r'ventas |otros \(|coste (materias|cafe|ingredientes|bebidas|cocteles)|'
    r'packaging)', re.I)
RX_INGRESO = re.compile(r'^(ventas|ingresos|otros\b)', re.I)

#: Qué partidas de la inversión son inmovilizado y con qué vida útil (§2.3.6,
#: TEC-20 y `NUEVO-02`: la base NO puede incluir circulante, stock ni
#: imprevistos).
AMORT_DEFECTO = {
    # ⚠️ T7/panadería (2026-08-29): «Amasadora», «Divisora + boleadora»,
    # «Laminadora» y «Salida humos + ventilación obrador» no casaban con
    # NINGÚN grupo y `_clasificar_amortizable` los mandaba por defecto a
    # 'no' — un obrador entero (18.000 € de horno aparte) quedándose fuera
    # de la base amortizable, la misma familia de defecto que TEC-20/
    # `NUEVO-02`, sólo que por FALTA de patrón en vez de por sobra. Se
    # añaden `humos`/`ventilaci[oó]n` a obra y el equipamiento específico de
    # panadería a maquinaria; son términos que no aparecen en ningún otro
    # hermano medido, así que la ampliación no reclasifica nada ajeno.
    'obra': (r'obra civil|adecuaci|reforma|instalaci|fontaner|el[eé]ctric|'
             r'climatizaci|extracci|proyecto t[eé]cnico|decoraci|interiorismo|'
             r'rotulaci|campana extractora|licencia de obras|humos|'
             r'ventilaci[oó]n',),
    'maquinaria': (r'equipamiento|maquina|m[aá]quina|horno|nevera|c[aá]mara|'
                   r'vitrina|mobiliario|barra|mostrador|tpv|vajilla|'
                   r'cristaler|cuberter|menaje|plancha|molinillo|lavavajillas|'
                   r'grifo|freidora|cafetera|batidora|tostadora|mesa|silla|'
                   r'taburete|estanter|fregadero|vinoteca|terraza|'
                   r'sandwichera|expositor|comandero|software|utensilios|'
                   r'amasadora|divisora|boleadora|laminadora|balanza',),
    # ⚠️ CRIT-02 — «Primera compra de despensa y cámaras» (6.500 €) y «Primera
    # compra de bodega y barra» (5.500 €) casaban con `c[aá]mara` y con
    # `barra` y se amortizaban a 10 años como si fueran maquinaria, inflando
    # la amortización en 1.200 €/año. Son EXISTENCIAS: su bloque se titula
    # «EXISTENCIAS INICIALES». Los hermanos las rotulan «Stock inicial …» y
    # por eso `stock` bastaba; el representante no. `primera compra` y
    # `existencias` van DELANTE porque `_clasificar_amortizable` prueba el
    # grupo «no» antes que los otros dos.
    'no': (r'fianza|primer mes|inmobiliaria|stock|existencias|primera compra|'
           r'fondo de maniobra|'
           r'colch[oó]n|imprevisto|marketing|lanzamiento|campa[ñn]a '
           r'lanzamiento|web|'
           r'constituci|notar[ií]a|registro|gestor[ií]a|seguro|licencia de '
           r'actividad|permiso|tasa|iva|marca|dise[ñn]o',),
}
#: Segunda red, por BLOQUE: nada de lo que cuelgue de un bloque de circulante
#: o de gasto es inmovilizado, se llame como se llame la partida. Cierra la
#: clase entera del defecto CRIT-02 para los hermanos de T7/T8, donde los
#: rótulos son otros. En el molde A-β (lista plana sin bloques) el nombre por
#: defecto es «INVERSIÓN» y no casa, así que no cambia nada allí.
RX_BLOQUE_NO_AMORT = re.compile(
    r'existencias|stock|g[ée]nero|fondo de maniobra|circulante|tesorer[ií]a|'
    r'marketing|lanzamiento|legales|constituci|gastos previos|'
    r'gastos de apertura', re.I)
# ⚠️ «campana extractora» convive con «Campaña lanzamiento RRSS» en el mismo
# libro y la primera va SIN tilde: por eso la lista de «no amortizable» exige
# la palabra «lanzamiento» detrás y la extractora se declara obra a mano. Es
# la trampa que `CLAUDE.md` documenta para el gate de ortografía.

# ==========================================================================
# Utilidades de escritura
# ==========================================================================


def ref(ws, coord):
    """Referencia con NOMBRE DE HOJA ACTUAL.

    Imprescindible: §1.7 renombra `'PyG 3 Anos'` → `'PyG 3 Años'` en
    `motor.cerrar()`, DESPUÉS de este grupo, y reescribe las referencias que
    encuentre. En la 2.ª pasada la hoja ya se llama con tilde y escribir el
    nombre viejo dejaría la fórmula apuntando a una hoja inexistente, que es
    justo lo que caza `gate_referencias`.
    """
    return "'" + ws.title + "'!" + coord


def _sin_comillas(formula):
    return re.sub(r'"[^"]*"', '', formula)


def fx(ws, coord, formula, fmt=None, align=None):
    """`motor.f()` con la guarda de §1.5 puesta de oficio.

    TODA fórmula del grupo va envuelta en `IFERROR(...,"")`, no sólo las
    divisiones. Motivo medido: con el libro en blanco el ticket es un número y
    el coste variable unitario es `""`, así que una simple RESTA devuelve
    `#¡VALOR!` y el error se propaga a la cifra estrella. Envolviéndolo todo,
    «sin dato» se escribe `""` (convención de familia) y el semáforo, que
    lleva la guarda `ISNUMBER` del §1.6, no pinta verde una celda vacía.

    Se hace aquí y no confiando en `motor.guardas()` porque el motor pasa por
    el libro ANTES que el grupo: lo que escribe el grupo ya no lo ve.
    """
    # RT-18 — «una referencia a una celda VACÍA devuelve 0, no ""». Un
    # espejo (`='0. Supuestos'!$B$24`) imprime «0 €» en un libro en blanco, y
    # la convención de la familia prohíbe expresamente escribir 0 donde no
    # hay dato: un libro vacío entregado así parece un plan con cifras.
    m = RX_ESPEJO.match(formula)
    if m:
        ref = m.group(1)
        formula = '=IF(' + ref + '="","",' + ref + ')'
    if not motor.RX_YA_GUARDADA.match(formula):
        formula = motor.iferror(formula)
    return motor.f(ws, coord, formula, fmt, align)


#: Fórmula que es un ESPEJO puro de otra celda: `='0. Supuestos'!$B$24`.
RX_ESPEJO = re.compile(r"^=((?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+)$")


def _limpiar_area(ws, r0, r1, ncols):
    """Vacía valores, estilos, combinadas, DV y CF de la zona de datos.

    Reconstruir es más seguro que insertar filas: `insert_rows` de openpyxl no
    mueve validaciones, formato condicional ni combinadas, y no reescribe
    fórmulas.
    """
    if r1 < r0:
        return
    for m in list(ws.merged_cells.ranges):
        cr = CellRange(str(m))
        if cr.max_row >= r0 and cr.min_row <= r1:
            _unmerge(ws, m)
    vacio = PatternFill()
    borde = Border()
    for r in range(r0, r1 + 1):
        for c in range(1, ncols + 1):
            cel = ws.cell(row=r, column=c)
            cel.value = None
            cel.fill = vacio
            cel.border = borde
            cel.font = Font()
            cel.alignment = Alignment()
            cel.number_format = 'General'
            cel.protection = Protection(locked=True)
            cel.hyperlink = None
        ws.row_dimensions[r].height = None
    _purgar_dv_area(ws, r0, r1)
    _purgar_cf_area(ws, r0, r1)
    ws._pl_cabeceras = None          # la cabecera cambia: invalida la caché
    ws._pl_editables = set()
    ws._pl_negativos = set()


def _rangos_fuera(sqref, r0, r1):
    fuera = []
    for r in str(sqref).split():
        try:
            cr = CellRange(r)
        except Exception:                                    # noqa: BLE001
            continue
        if cr.max_row >= r0 and cr.min_row <= r1:
            continue
        fuera.append(str(cr))
    return fuera


def _purgar_dv_area(ws, r0, r1):
    quedan = []
    for dv in ws.data_validations.dataValidation:
        restos = _rangos_fuera(dv.sqref, r0, r1)
        if not restos:
            continue
        dv.sqref = ' '.join(restos)
        quedan.append(dv)
    ws.data_validations.dataValidation = quedan


def _purgar_cf_area(ws, r0, r1):
    from openpyxl.formatting.formatting import ConditionalFormattingList
    supervivientes = []
    for cf in ws.conditional_formatting:
        restos = _rangos_fuera(cf.sqref, r0, r1)
        if not restos:
            continue
        supervivientes.append((' '.join(restos), list(cf.rules)))
    nueva = ConditionalFormattingList()
    for sqref, reglas in supervivientes:
        for r in reglas:
            nueva.add(sqref, r)
    ws.conditional_formatting = nueva


def _cabecera(ws):
    """Fila de cabecera de la tabla (la que tiene 3+ rótulos)."""
    fila, _ = motor._fila_cabecera(ws)
    return fila or 4


def _es_mayusculas(texto):
    letras = [c for c in motor.norm(texto) if c.isalpha()]
    if not letras:
        return False
    return str(texto).upper() == str(texto)


def _unmerge(ws, m):
    """`unmerge_cells` revienta con KeyError cuando la combinada quedó
    colgando de una fila BORRADA (`delete_rows` no toca `merged_cells`).
    Se quita entonces del registro a mano."""
    try:
        ws.unmerge_cells(str(m))
    except KeyError:
        try:
            ws.merged_cells.ranges.remove(m)
        except Exception:                                    # noqa: BLE001
            pass


def _borrar_fila(ws, r):
    """`delete_rows` + arreglo de las COMBINADAS, que openpyxl no mueve.

    Medido: al borrar la fila duplicada del RGSEAA, las seis cabeceras de
    fase (combinadas `A4:F4`, `A15:F15`…) se quedaban en su posición vieja.
    Las filas de contenido que subían a esas posiciones caían DENTRO de una
    combinada muerta y openpyxl las serializa VACÍAS: se perdían «Reforma y
    acondicionamiento del local» y «Publicar ofertas de empleo» sin que nada
    lo dijera.
    """
    from openpyxl.worksheet.cell_range import MultiCellRange
    rangos = []
    for m in list(ws.merged_cells.ranges):
        cr = CellRange(str(m))
        if cr.min_row <= r <= cr.max_row and cr.min_row == cr.max_row:
            continue                       # la combinada de la fila borrada
        if cr.min_row > r:
            cr.shift(row_shift=-1)
        elif cr.max_row >= r:
            cr.max_row -= 1
        rangos.append(str(cr))
    ws.delete_rows(r)
    ws.merged_cells = MultiCellRange()
    for ref in rangos:
        ws.merged_cells.add(ref)
    alturas = sorted((f, d.height) for f, d in ws.row_dimensions.items()
                     if d.height)
    for fila_, alto in alturas:
        if fila_ > r:
            ws.row_dimensions[fila_ - 1].height = alto
    if alturas and alturas[-1][0] > r:
        ws.row_dimensions[alturas[-1][0]].height = None


def _desmerge_fila(ws, fila):
    """Deshace cualquier combinada que toque la fila: `MergedCell.value` es
    de sólo lectura y escribir en ella revienta con AttributeError."""
    for m in list(ws.merged_cells.ranges):
        cr = CellRange(str(m))
        if cr.min_row <= fila <= cr.max_row:
            _unmerge(ws, m)


def _ancho_combinado(ws, fila):
    for m in ws.merged_cells.ranges:
        cr = CellRange(str(m))
        if cr.min_row <= fila <= cr.max_row:
            return cr.max_col
    return 1


def _pie(ws, r0):
    """Filas de pie («ChefBusiness.co — …», «NOTA: …») para reponerlas."""
    fuera = []
    for r in range(r0, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and re.match(r'^(NOTA|ChefBusiness|\*)', v):
            fuera.append(v)
    return fuera


class Rejilla(object):
    """Constructor de filas en dos fases.

    Primero se declaran todas las filas (con las fórmulas como funciones que
    reciben el resolvedor de coordenadas) y después se escriben: así una fila
    puede referenciar a otra que todavía no existía cuando se declaró, sin que
    el módulo tenga que saber números de fila.
    """

    def __init__(self, ws, fila0, driver=None, cols_driver=('B', 'C', 'D')):
        self.ws = ws
        self.fila0 = fila0
        self.filas = []
        self._pos = {}
        #: RT-18 — «celda con el DRIVER vacío» se escribe `""`, nunca 0. El
        #: driver es la celda de la que cuelga toda la fila (los ingresos del
        #: año, el bruto del puesto…): si está vacía, la fila entera queda en
        #: blanco en vez de imprimir «0 €». Firma: `(rejilla, columna, clave)`.
        self.driver = driver
        self.cols_driver = cols_driver

    def add(self, clave=None, **kw):
        self.filas.append(dict(kw, clave=clave))
        if clave:
            self._pos[clave] = self.fila0 + len(self.filas) - 1
        return self.fila0 + len(self.filas) - 1

    def fila(self, clave):
        return self._pos[clave]

    def c(self, clave, col='B', absoluta=False):
        fila = self._pos[clave]
        if absoluta:
            return '$' + col + '$' + str(fila)
        return col + str(fila)

    def r(self, clave, col='B'):
        """Referencia con hoja: para que la use otra hoja del libro."""
        return ref(self.ws, self.c(clave, col, absoluta=True))

    @property
    def ultima(self):
        return self.fila0 + len(self.filas) - 1


#: RT-24 — 38 notas escritas como `=IFERROR("texto literal","")`. Una
#: constante de texto no puede provocar un error: la guarda no protege de
#: nada, convierte un comentario en fórmula, engorda el recuento de «fórmulas
#: nuevas» que sostiene la promesa comercial y obliga a `inject_cache` a
#: cachear texto. RT-17: lo mismo con `=IFERROR(0,"")` y `=IFERROR(1,"")`.
#: Se detectan aquí, en el volcado, y se escriben como VALOR. Las notas que
#: SÍ interpolan una celda con `TEXT()` no casan con estos patrones y siguen
#: siendo fórmula, que es lo que las mantiene al día.
RX_SOLO_TEXTO = re.compile(r'^="((?:[^"]|"")*)"$')
RX_SOLO_NUMERO = re.compile(r'^=(-?\d+(?:\.\d+)?)$')


def escribir(rej, cols_texto=('A',)):
    """Vuelca las filas declaradas. `formulas`/`valores` por columna."""
    ws = rej.ws
    for i, spec in enumerate(rej.filas):
        fila = rej.fila0 + i
        rot = spec.get('rot')
        if rot is not None:
            motor.val(ws, 'A' + str(fila), rot,
                      bold=spec.get('bold', False),
                      wrap=spec.get('wrap'))
        verdes = spec.get('verdes')
        for col, valor in sorted((spec.get('valores') or {}).items()):
            editable = (col in verdes) if verdes is not None \
                else spec.get('verde', False)
            motor.val(ws, col + str(fila), valor, spec.get('fmt_' + col)
                      or spec.get('fmt'), verde_=editable,
                      bold=spec.get('bold', False))
        for col, fabrica in sorted((spec.get('formulas') or {}).items()):
            formula = fabrica(rej) if callable(fabrica) else fabrica
            if formula is None:
                continue
            if not str(formula).startswith('='):
                # una nota declarada en `formulas` es un VALOR: envolverla en
                # `IFERROR(...)` produciría una fórmula rota
                motor.val(ws, col + str(fila), formula,
                          bold=spec.get('bold', False))
                continue
            m_txt = RX_SOLO_TEXTO.match(formula)
            m_num = RX_SOLO_NUMERO.match(formula)
            if m_txt is not None:
                motor.val(ws, col + str(fila), m_txt.group(1).replace('""',
                                                                      '"'),
                          bold=spec.get('bold', False))
            elif m_num is not None:
                valor = float(m_num.group(1))
                motor.val(ws, col + str(fila),
                          int(valor) if valor == int(valor) else valor,
                          spec.get('fmt_' + col) or spec.get('fmt'),
                          bold=spec.get('bold', False))
            else:
                drv = rej.driver
                if drv is not None and col in rej.cols_driver \
                        and not spec.get('sin_driver'):
                    ref = drv(rej, col, spec.get('clave'))
                    # ⚠️ B14 / REF-22 / REF-18 (2026-09-05): la fila «Punto de
                    # equilibrio alcanzado» ya trae su propia guarda con la
                    # MISMA celda, y el volcado la envolvía otra vez:
                    # `IF(B9="","",IF(B9="","",…))`. No cambia el resultado —
                    # por eso ni la idempotencia ni el `data_only` lo veían—
                    # pero en una celda donde la guarda interna no fuese
                    # idéntica a la externa el segundo envoltorio se comería
                    # el operando bueno.
                    guarda = 'IF(' + ref + '="","",' if ref else None
                    if ref and not formula[1:].startswith(guarda):
                        formula = '=' + guarda + formula[1:] + ')'
                fx(ws, col + str(fila), formula,
                   spec.get('fmt_' + col) or spec.get('fmt'))
                if spec.get('bold'):
                    cel = ws[col + str(fila)]
                    cel.font = Font(bold=True, size=cel.font.size)
        if spec.get('alto'):
            ws.row_dimensions[fila].height = spec['alto']


def texto_pct(refcelda, prefijo, sufijo=''):
    """Frase con el porcentaje GENERADO desde la celda (§7-bis.11).

    Se usa `"0%"` a propósito: `TEXT()` con decimales imprime el separador
    anglosajón en el valor cacheado y este libro ya arrastraba el defecto de
    mezclar «57.9» con «10,73» en la misma frase (TEC-21).
    """
    return ('="' + prefijo + '"&TEXT(' + refcelda + ',"0%")&"' + sufijo + '"')


def texto_num(refcelda, prefijo, sufijo='', fmt='0'):
    return ('="' + prefijo + '"&TEXT(' + refcelda + ',"' + fmt + '")&"'
            + sufijo + '"')


def dec1(refcelda):
    """Un decimal con COMA, idéntico en Excel y en la caché de pycel.

    `TEXT(x,"0.0")` no vale: el separador decimal de un código de formato se
    guarda siempre en canónico anglosajón, así que Excel en español pinta
    «1,4» y la caché que ve el cliente en Vista previa, en el móvil o en
    Google Sheets dice «1.4». En la MISMA frase que habla de «1,8
    rotaciones»: es exactamente la mezcla de separadores que cerró TEC-21.

    Y `"0,0"` es peor todavía: la coma de un código de formato es el
    separador de MILES, nunca el decimal, por mucho que la interfaz española
    se teclee así. Imprimía literalmente «Son 01 servicios por plaza y día».

    Se compone la cifra a mano, con la coma como TEXTO. El `ROUND` va FUERA
    del troceado para que el acarreo sea correcto: con 1,96 hay que leer
    «2,0», no «1,10».
    """
    r = 'ROUND(' + refcelda + ',1)'
    return ('TEXT(INT(' + r + '),"0")&","&TEXT(ROUND(MOD(' + r
            + ',1)*10,0),"0")')


# ==========================================================================
# Lectura del fichero de partida
# ==========================================================================


#: Paréntesis que sólo contienen un parámetro numérico: se van del rótulo
#: porque mienten en cuanto se toca la celda (§7-bis.11). El dato vuelve en la
#: columna de notas, generado con `TEXT()`.
#: ⚠️ REF-06 (`auditorias/planes-v2-hermano-plan-negocio-tapas-bar-ref.json`,
#: 2026-08-29): la unidad `%` es un carácter NO-word y el paréntesis de cierre
#: `)` también, así que `\b` entre los dos NUNCA casa (no-word → no-word no es
#: frontera de palabra) — «Imprevistos (8%)» sobrevivía intacto y la regla
#: `INVERSION['imprevistos'] = ('suprimir', …)` no encontraba su clave porque
#: `motor.norm(rot)` seguía trayendo el paréntesis. El `\b` sólo hace falta
#: DETRÁS de las alternativas que son letras (`años`/`meses`/`mes`); en `%` se
#: quita.
RX_PARENTESIS_NUM = re.compile(
    r'\s*\((?:incl\.?\s*)?[^()]*?\d+(?:[.,]\d+)?\s*'
    r'(?:%|a[ñn]os?\b|meses\b|mes\b)[^()]*\)', re.I)


def _limpiar_rotulo(rotulo):
    """«Imprevistos (8%)» → «Imprevistos». Devuelve (rótulo, cambió)."""
    nuevo = RX_PARENTESIS_NUM.sub('', rotulo).strip(' -—:')
    if nuevo and nuevo != rotulo:
        return nuevo, True
    return rotulo, False


def _seccion(ws, cab, rx_inicio, rx_fin):
    """Rango de filas de una SECCIÓN, acotado por sus rótulos.

    Buscar dentro de la hoja entera no falla: acierta en el sitio equivocado.
    Sin acotar, el barrido de «costes fijos» del P&L de A-β se llevaba por
    delante el EBITDA, el impuesto y el resultado neto, que están debajo y
    también tienen número. (Misma lección que los regex de sección del blog.)
    """
    ini = fin = None
    for r in range(cab, ws.max_row + 1):
        rot = motor._rotulo_de_fila(ws, r, max_col=1)
        if not rot:
            continue
        if ini is None:
            if re.search(rx_inicio, rot, re.I):
                ini = r + 1
            continue
        if re.search(rx_fin, rot, re.I):
            fin = r - 1
            break
    if ini is None:
        return None, None
    return ini, (fin if fin is not None else ws.max_row)


def _partidas(ws, cab, col_importe=2, canon=None, desde=None, hasta=None):
    """Partidas de una tabla: (fila, rótulo, importe, nota, bloque).

    Distingue encabezado de bloque (texto en mayúsculas sin importe **o con
    importe de fórmula**, que es como quedan tras el subtotal de §2.2) de
    partida (rótulo + número) y de pie. Funciona igual sobre el fichero
    original y sobre el reconstruido: de eso depende la idempotencia.
    """
    bloque = None
    fuera = []
    for r in range(desde or (cab + 1), (hasta or ws.max_row) + 1):
        rot = motor._rotulo_de_fila(ws, r, max_col=1)
        if not rot:
            continue
        v = ws.cell(row=r, column=col_importe).value
        if _es_mayusculas(rot) and not motor._es_numero(v):
            if not motor.RX_TOTAL.match(rot):
                bloque = rot
            continue
        if motor.RX_TOTAL.match(rot) or re.match(r'^(NOTA|ChefBusiness|\*)',
                                                 rot):
            continue
        if canon is not None and canon.search(rot):
            continue
        if not motor._es_numero(v):
            continue
        nota = None
        for c in range(col_importe + 1, min(ws.max_column, 6) + 1):
            vv = ws.cell(row=r, column=c).value
            # la columna de marca de IVA no es una nota: leerla como tal la
            # copiaba a la columna de notas en la 2.ª pasada y rompía la
            # idempotencia (RD-14)
            if re.search(r'iva', motor.cabecera_de_columna(ws, c) or '',
                         re.I):
                continue
            if isinstance(vv, str) and vv.strip() \
                    and not vv.startswith('=') and '%' not in (
                        ws.cell(row=r, column=c).number_format or ''):
                nota = vv.strip()
                break
        fuera.append((r, rot, float(v), nota, bloque))
    return fuera


#: RD-14 / RT-25 — clasificación por defecto del IVA de cada partida. Se
#: marca «No» sólo lo que la ley deja fuera: la fianza (garantía, no entrega
#: de bienes), el colchón de caja, los seguros (exentos, art. 20.Uno.16 LIVA)
#: y las tasas, licencias y aranceles registrales (no sujetos). Todo lo demás
#: nace en «Sí» y el comprador lo corrige en su desplegable.
RX_SIN_IVA = re.compile(
    r'^(fianza|fondo de maniobra|colch[oó]n|seguros?\b|licencias?\b|'
    r'permisos?\b|tasas?\b|impuesto|registro mercantil|'
    r'constituci[oó]n)', re.I)


def _lleva_iva(rotulo):
    return 'No' if RX_SIN_IVA.search(rotulo or '') else 'Sí'


def _clasificar_amortizable(rotulo, tablas):
    for grupo in ('no', 'obra', 'maquinaria'):
        for rx in tablas.get(grupo, ()):
            if re.search(rx, rotulo, re.I):
                return grupo
    return 'no'


def compilar_vocabulario(voc):
    """(regex, mapa) del vocabulario del oficio (M9), o `(None, {})`."""
    mapa = dict((motor.norm(k), v) for k, v in (voc or {}).items() if k and v)
    if not mapa:
        return None, {}
    rx = re.compile(r'\b(' + '|'.join(
        re.escape(k) for k in sorted(mapa, key=len, reverse=True))
        + r')\b', re.I | re.U)
    return rx, mapa


def traducir(texto, rx, mapa):
    """Palabras completas, respetando MAYÚSCULAS y mayúscula inicial (M9)."""
    if not rx or not isinstance(texto, str):
        return texto
    return rx.sub(lambda m: motor._reponer_caso(m.group(0),
                                                mapa[motor.norm(m.group(0))]),
                  texto)


def _enumerar_base(rotulos, claves, primero, ultimo, rej):
    """M4 — los rótulos REALES que entran en la base de los imprevistos.

    `claves` son las partidas clasificadas como 'obra' dentro del bloque; si
    el molde no tiene ninguna, la fórmula suma el bloque entero y la
    enumeración tiene que decir eso mismo.
    """
    mapa = dict(rotulos)
    if claves:
        orden = list(claves)
    else:
        f0, f1 = rej.fila(primero), rej.fila(ultimo)
        orden = [k for k, _r in rotulos
                 if k in mapa and f0 <= rej.fila(k) <= f1]
    vistos, limpio = set(), []
    for k in orden:
        rot = str(mapa.get(k) or '').replace('"', '').strip()
        if not rot or motor.norm(rot) in vistos:
            continue
        vistos.add(motor.norm(rot))
        limpio.append(rot)
    if not limpio:
        return 'las partidas de obra de este bloque'
    if len(limpio) == 1:
        return limpio[0]
    return ', '.join(limpio[:-1]) + ' y ' + limpio[-1]


def _num(valor, defecto=None):
    if motor._es_numero(valor):
        return float(valor)
    return defecto


# ==========================================================================
# El modelo del producto
# ==========================================================================
class Plan(object):

    def __init__(self, wb, det, pid, params, contenido, cambios):
        self.wb = wb
        self.det = det
        self.pid = pid
        self.p = params
        self.c = contenido
        self.cambios = cambios
        self.molde = det['molde']
        self.numerado = self.molde == 'A-alfa'
        self.concepto = self.dato('CONCEPTO', pid)
        self.rej = {}
        #: Rejillas declaradas y pendientes de volcar. El volcado va al FINAL
        #: de todo (§ ver `post`): las hojas se citan entre sí en las dos
        #: direcciones —el P&L lee los intereses de Financiación y el fondo de
        #: maniobra de Inversión sale de los costes fijos del P&L—, así que
        #: ninguna fórmula puede resolverse hasta que estén todas colocadas.
        self.pendientes = []
        if det['tipo'] != 'plan_financiero':
            # el checklist comparte módulo de contenido pero no tiene ni
            # supuestos ni hojas de modelo: se construye sin ellas
            return
        for clave, nombres in HOJAS.items():
            ws = None
            for nombre in nombres:
                ws = motor.hoja(wb, nombre)
                if ws is not None:
                    break
            setattr(self, 'ws_' + clave, ws)
        self.ws_sup = motor.hoja(wb, motor.HOJA_SUPUESTOS, obligatoria=True)
        self.ws_ins = motor.hoja(wb, 'Instrucciones', obligatoria=True)
        for clave, nombre, prefijo in NUEVAS:
            titulo = (prefijo if self.numerado else '') + nombre
            ws = motor.hoja(wb, titulo) or motor.hoja(wb, nombre)
            if ws is None:
                ws = wb.create_sheet(titulo)
            setattr(self, 'ws_' + clave, ws)

    # -- acceso al módulo de contenido -----------------------------------
    def dato(self, clave, defecto=None):
        if self.c is None:
            return defecto
        valor = getattr(self.c, clave, None)
        if valor is None and isinstance(getattr(self.c, 'CONTENIDO', None),
                                        dict):
            valor = self.c.CONTENIDO.get(clave)
        return defecto if valor is None else valor

    def anota(self, texto):
        self.cambios.append(texto)

    # -- §M9: vocabulario del oficio --------------------------------------
    def vocabulario(self):
        """M9 / R22-PAN-11 / R22-CAF-20 / REF-17 — el driver por su nombre.

        El motor rotula «Cubiertos/día», «Coste variable por cubierto» y «en
        sala» en los cinco hermanos, y sólo dos tienen mesas: la panadería
        cuenta TRANSACCIONES DE MOSTRADOR y la cafetería y el food truck
        cuentan CLIENTES — y así lo dicen todas sus notas, que contradecían a
        su propio rótulo en la misma fila. El módulo de contenido declara
        `VOCABULARIO = {'cubierto': 'transacción', …}` y aquí se hace UNA
        pasada final sobre las hojas que construye este grupo.

        Reglas: sólo palabras completas (`\b`), se respeta la mayúscula
        inicial y las MAYÚSCULAS completas, no se tocan los nombres de hoja
        (renombrarlos rompería todas las referencias) y, dentro de una
        fórmula, sólo se traduce lo que va entre comillas dobles: las
        referencias a hoja van entre comillas SIMPLES y no se rozan.
        """
        rx, mapa = compilar_vocabulario(self.dato('VOCABULARIO', {}))
        if not rx:
            return 0

        def _tr(txt):
            return traducir(txt, rx, mapa)

        def _tr_formula(txt):
            # sólo los literales de texto: `="… cubiertos …"&TEXT(B9,"0")`
            trozos = txt.split('"')
            for i in range(1, len(trozos), 2):
                trozos[i] = _tr(trozos[i])
            return '"'.join(trozos)

        hojas = [getattr(self, 'ws_' + c, None) for c in
                 ('sup', 'pyg', 'inversion', 'personal', 'equilibrio',
                  'escenarios', 'tesoreria', 'financiacion', 'ins')]
        n = 0
        for ws in hojas:
            if ws is None:
                continue
            for row in ws.iter_rows():
                for cel in row:
                    v = cel.value
                    if not isinstance(v, str) or not v.strip():
                        continue
                    nuevo = (_tr_formula(v) if v.startswith('=') else _tr(v))
                    if nuevo != v:
                        cel.value = nuevo
                        n += 1
        self.anota('Vocabulario del oficio aplicado en ' + str(n)
                   + ' celdas (' + ', '.join(
                       sorted(k + '→' + mapa[k] for k in mapa)) + ') — M9')
        return n

    # -- §A6: rótulos de filas canónicas y preservadas --------------------
    def _reglas(self, clave):
        """`INVERSION`/`FIJOS` con ALIAS por el rótulo NUEVO (A6 / M-05/M-09).

        La regla admite ahora una tupla de 3: `(importe|None, nota,
        rotulo_nuevo)`. La clave del diccionario es el rótulo VIEJO, que es el
        que trae el fichero de partida; en la 2.ª pasada la hoja ya lleva el
        NUEVO y la búsqueda por rótulo no lo encontraría, así que el rótulo
        nuevo se indexa también: sin este alias el nombre revertiría en la
        segunda pasada y la idempotencia se rompería.
        """
        cache = getattr(self, '_reglas_cache', None)
        if cache is None:
            cache = self._reglas_cache = {}
        if clave in cache:
            return cache[clave]
        reglas = dict(self.dato(clave, {}) or {})
        for _v in list(reglas.values()):
            if len(_v) > 2 and _v[2]:
                reglas.setdefault(motor.norm(_v[2]), _v)
        cache[clave] = reglas
        return reglas

    def _rotulo_regla(self, clave, rot):
        """Rótulo definitivo de una fila canónica según `INVERSION`/`FIJOS`."""
        accion = self._reglas(clave).get(motor.norm(rot))
        if accion and len(accion) > 2 and accion[2]:
            return accion[2]
        return rot

    # -- §2.1 -------------------------------------------------------------
    def supuestos_altas(self):
        """`0. Supuestos`: da de alta los parámetros y sus refs (§2.1).

        Va ANTES que ninguna otra hoja: todas cablean contra `p.ref(clave)`.
        """
        ws = self.ws_sup
        motor.val(ws, 'A1', 'SUPUESTOS — aquí se teclean las TASAS y los '
                  'DRIVERS del modelo', bold=True)
        motor.val(ws, 'A2', 'Cambia las celdas VERDES: el resto del libro se '
                  'recalcula solo. Las partidas de gasto se teclean en su '
                  'hoja (también en verde) y no se repiten aquí.', wrap=True)
        for coord, texto in BLOQUES_EXTRA:
            motor.val(ws, coord, texto, bold=True)
        propios = self.dato('SUPUESTOS', {}) or {}
        leidos = self._leer_drivers()
        for clave, coord, etiqueta, defecto, fmt, nota in SUPUESTOS_BASE:
            fuente = 'valor por defecto del grupo A'
            valor = defecto
            if clave in leidos and leidos[clave] is not None:
                valor, fuente = leidos[clave], 'leído del fichero original'
            if clave in propios:
                spec = propios[clave]
                etiqueta = spec[1] or etiqueta
                valor = spec[2] if spec[2] is not None else valor
                fmt = spec[3] or fmt
                nota = spec[4] or nota
                fuente = spec[5] or 'contenido del producto'
                coord = spec[0] or coord
            if valor is None:
                valor = 0
            self.p.alta(clave, etiqueta, valor, fmt, nota, coord=coord)
            self.cambios.append('Supuestos!' + coord + ' ' + clave + ' = '
                                + str(valor) + ' (' + fuente + ')')
        motor.anchos(ws, {'A': 48, 'B': 16, 'C': 74})

    def supuestos_calculadas(self):
        """Las DOS celdas de `0. Supuestos` que no se teclean (§2.1).

        `pct_comida` sale de sumar el peso de las líneas de comida del P&L y
        `pct_bebida` es el resto: así el mix nunca puede sumar distinto de
        100 % y la mezcla vive en un solo sitio. Se escriben al final porque
        necesitan las coordenadas del P&L, que se declara después.
        """
        ws = self.ws_sup
        rej = self.rej['pyg']
        lineas = self.lineas_ingreso()
        comida = [i for i, l in enumerate(lineas) if l[2] == 'comida']
        if getattr(self, 'mix_en_supuestos', False):
            # RD-23 / RC-06 / RT-13 — la dependencia INVERTIDA: el mix se
            # teclea aquí, que es donde la hoja de Instrucciones dice que
            # está, y el P&L lo lee. Antes era al revés: B11 era una fórmula
            # BLOQUEADA que leía el P&L, y el cliente que seguía las
            # instrucciones al pie de la letra no podía cambiar el segundo
            # driver del libro.
            peso = lineas[comida[0]][1] if comida else 0.65
            actual = ws['B11'].value
            if not motor._es_numero(actual) or float(actual) <= 0:
                motor.val(ws, 'B11', peso, motor.FMT_PCT, verde_=True)
            else:
                ws['B11'].number_format = motor.FMT_PCT
                motor.verde(ws, 'B11')
            motor.val(ws, 'C11', 'Peso de la COMIDA sobre el total de ventas. '
                      'Es un INPUT: cámbialo aquí y el P&L, la tesorería y el '
                      'IVA repercutido se recalculan solos')
            # B1 / R-08 / REF-15 — sin la guarda, un libro VACIADO imprime
            # «100,0 %» de bebida (`1-""` = 1): un falso dato, no un blanco.
            fx(ws, 'B12', '=IF(COUNT(B11)=0,"",1-B11)', motor.FMT_PCT)
            motor.val(ws, 'C12', 'Se calcula como el resto: comida + bebida = '
                      '100 %')
            ws['B12'].fill = PatternFill()
            ws['B12'].protection = Protection(locked=True)
        else:
            suma = ('+'.join(rej.r('lin_%d' % i, 'E') for i in comida)
                    if comida else '0')
            fx(ws, 'B11', '=' + suma, motor.FMT_PCT)
            motor.val(ws, 'C11', 'Suma del peso de las líneas de COMIDA del '
                      'P&L: el mix vive en un solo sitio')
            fx(ws, 'B12', '=IF(COUNT(B11)=0,"",1-B11)', motor.FMT_PCT)
            motor.val(ws, 'C12', 'Se calcula como el resto: comida + bebida = '
                      '100 %')
            for coord in ('B11', 'B12'):
                ws[coord].fill = PatternFill()
                ws[coord].protection = Protection(locked=True)
        # nota generada: el PVP con IVA equivalente al ticket sin IVA (TEC-11)
        # A3 / REF-04 — el PVP ya no se compone con dos porcentajes globales
        # (comida al reducido y bebida a su mezcla): lo compone el IVA MEDIO
        # PONDERADO de las líneas de venta del P&L, que es donde cada línea
        # declara su tipo. Para el representante el resultado es idéntico
        # (dos líneas, comida al reducido y bebida a la mezcla RD-17); para
        # una panadería con pan común al 4 %, no.
        tic = self._loc('ticket_medio')
        iva_med = self.iva_ventas
        motor.val(ws, 'A9', 'PVP equivalente con IVA (calculado)')
        fx(ws, 'B9', '=IF(' + tic + '="","",' + tic + '*(1+' + iva_med + '))',
           motor.FMT_EUR)
        fx(ws, 'C9', '="Ticket sin IVA más el IVA medio ponderado de tus '
           'líneas de venta: "&TEXT(' + iva_med + '*100,"0")&" %. En sala '
           'todo va al 10 % (art. 91.Uno.2.2.º de la Ley 37/1992, alcohol '
           'incluido); el alcohol que sale por delivery y el pan común, si '
           'los tienes, mueven la media. Aquí va redondeado. Compáralo con '
           'el rango del sector, que va con IVA."')
        # RD-24 — la rotación implícita, calculada desde el aforo en celda,
        # en vez de una nota que citaba mal su fuente (presentaba el 1,8 del
        # documento como un TECHO cuando ahí es un mínimo exigible, y usaba
        # 56 plazas contra las 40-50 del propio documento).
        aforo = self._loc('aforo')
        rota = self.dato('ROTACION', {}) or {}
        if not rota.get('activa', True):
            # A5 — el bloque se APAGA entero: se escriben las tres celdas a
            # vacío (no basta con no escribirlas: el fichero de partida trae
            # las de la versión anterior y sobrevivirían).
            for coord in ('A50', 'B50', 'C50'):
                motor.val(ws, coord, None)
                ws[coord].number_format = 'General'
        else:
            # ⚠️ MOT-02 — la frase de referencia («MÍNIMO de 1,8 rotaciones
            # por mesa y servicio») es del documento del REPRESENTANTE. En
            # tapas-bar el docx dice otra cosa (1,8 en almuerzo y 1,5 en
            # cena) y la celda comparaba ese umbral con rotaciones AL DÍA:
            # medía una cosa y el número otra. Ahora el molde trae la suya.
            referencia = (rota.get('referencia')
                          or 'El documento de este plan pide un MÍNIMO de '
                             '1,8 rotaciones por mesa y servicio en '
                             'temporada alta: compáralo con tu horario real '
                             'antes de darlo por bueno.')
            motor.val(ws, 'A50', 'Rotaciones al día implícitas (calculado)')
            fx(ws, 'B50', '=' + self._loc('cubiertos_dia') + '/' + aforo,
               motor.FMT_DEC2)
            # ⚠️ este era el único TEXT() del libro con decimales y el único
            # que seguía mezclando separadores dentro de la frase (ver
            # `dec1`).
            fx(ws, 'C50', '="Son "&' + dec1('B50')
               + '&" servicios por plaza y día '
               'sobre las "&TEXT(' + aforo + ',"0")&" plazas de la celda de '
               'arriba. ' + str(referencia).replace('"', '') + '"')
        motor.anchos(ws, {'A': 48, 'B': 16, 'C': 74})
        motor.print_setup(ws)

    def _loc(self, clave):
        """Coordenada de un parámetro DENTRO de `0. Supuestos` (sin hoja)."""
        return self.p.ref(clave).split('!')[-1]

    def _leer_drivers(self):
        """Drivers que el fichero ya trae, para no inventarlos (§1.3).

        En A-α están en el P&L; en A-β viven repartidos entre `Escenarios`
        (clientes/día, ticket, días) y `Punto Equilibrio` — con DOS calendarios
        distintos, que es `NUEVO-03`. Se toma el de `Escenarios`, que es el que
        declara días/año.
        """
        fuera = {}

        def por_rotulo(ws, patron, col='B'):
            if ws is None:
                return None
            for r in range(1, ws.max_row + 1):
                rot = motor._rotulo_de_fila(ws, r, max_col=1)
                if rot and re.search(patron, rot, re.I):
                    return _num(ws[col + str(r)].value)
            return None

        fuera['cubiertos_dia'] = por_rotulo(
            self.ws_pyg, r'^cubiertos') or por_rotulo(
                self.ws_escenarios, r'^(clientes|cubiertos)', 'C')
        fuera['ticket_medio'] = por_rotulo(
            self.ws_pyg, r'^ticket') or por_rotulo(
                self.ws_escenarios, r'^ticket', 'C')
        fuera['dias_apertura'] = por_rotulo(
            self.ws_pyg, r'^d[ií]as') or por_rotulo(
                self.ws_escenarios, r'd[ií]as', 'C')
        fuera['alquiler_mes'] = None
        alq = por_rotulo(self.ws_pyg, r'^alquiler')
        if alq:
            fuera['alquiler_mes'] = round(alq / 12.0, 2)
        sum_ = por_rotulo(self.ws_pyg, r'^suministros')
        if sum_:
            fuera['suministros_mes'] = round(sum_ / 12.0, 2)
        seg = por_rotulo(self.ws_pyg, r'^seguros?')
        if seg:
            fuera['seguros_ano'] = seg
        return fuera

    # -- líneas de ingreso -------------------------------------------------
    def lineas_ingreso(self):
        """Las líneas de venta y su peso. Une A-α (1 línea) con A-β (4)."""
        if getattr(self, '_lineas', None) is not None:
            return self._lineas
        propias = self.dato('LINEAS_INGRESO')
        if propias:
            self._lineas = [tuple(l) for l in propias]
            return self._lineas
        ws = self.ws_pyg
        cab = _cabecera(ws)
        # ACOTAR la sección antes de buscar dentro: fuera del bloque de
        # ingresos hay más filas que empiezan por «Ventas» o «Otros».
        r0, r1 = _seccion(ws, cab, r'^ingresos$|^ingresos\b',
                          r'^(total ingresos|ingresos brutos|ingresos '
                          r'totales|costes variables)')
        total = None
        lineas = []
        for r in range(r0 or cab + 1, (r1 or ws.max_row) + 1):
            rot = motor._rotulo_de_fila(ws, r, max_col=1)
            if not rot:
                continue
            if motor.RX_TOTAL.match(rot) or _es_mayusculas(rot):
                v = _num(ws.cell(row=r, column=2).value)
                if v and re.search(r'ingreso', rot, re.I):
                    total = v
                continue
            if not RX_INGRESO.match(rot):
                continue
            peso = _num(ws.cell(row=r, column=5).value)
            importe = _num(ws.cell(row=r, column=2).value)
            lineas.append([rot, peso, importe])
        if not lineas:
            self._lineas = [('Ingresos de comida', 0.65, 'comida', None,
                             'reparto por defecto del grupo A'),
                            ('Ingresos de bebida', 0.35, 'bebida', None,
                             'el resto')]
            return self._lineas
        suma = sum(l[2] for l in lineas if l[2]) or 1.0
        fuera = []
        for rot, peso, importe in lineas:
            if peso is None:
                peso = round((importe or 0) / (total or suma), 4)
            grupo = 'bebida' if re.search(
                r'bebida|c[oó]ctel|caf[eé]|vino|cerveza|refresco|zumo|'
                r'destilado|barra', rot, re.I) else 'comida'
            fuera.append((rot, peso, grupo, None, 'leído del fichero'))
        self._lineas = fuera
        return self._lineas

    # -- §2.6 -------------------------------------------------------------
    def personal(self):
        """`Personal`: SS en celda, fila cuadrada y TOTAL cerrado (§2.6)."""
        ws = self.ws_personal
        cab = _cabecera(ws)
        # B2 / R-11 / REF-16 — el SUBTÍTULO de la hoja llevaba el tipo de
        # Seguridad Social escrito a mano («33.4 %») y desmentía a la celda de
        # la que cuelga toda la columna E ('0. Supuestos'!B20 = 33 %). Vive
        # POR ENCIMA de la cabecera, así que ni `_limpiar_area` (que arranca
        # en `cab`) ni el filtro del pie lo tocaban. Pasa a fórmula, igual que
        # ya se hace en 'Inversión Inicial'!D30/D32/D34. Se escribe «0» y no
        # «0,0%» a propósito: la coma de un código de formato es el separador
        # de MILES y el punto depende del locale (misma trampa que `dec1`).
        # ⚠️ M13 / R22-PAN-17 (2026-09-05). B2 del parche 2.2 sólo REESCRIBÍA
        # el subtítulo cuando ya había uno con el porcentaje a mano, así que
        # la panadería —cuya fila 2 viene VACÍA del fichero v1.1— se quedó
        # sin él y el dry-run salió 13/13 verde igualmente: es la única hoja
        # del libro sin subtítulo, y en ella el lector no ve de un vistazo
        # las 14 pagas ni el 33 % de Seguridad Social con los que se calcula
        # toda la columna. Ahora, si no hay ninguno arriba de la cabecera, se
        # ESCRIBE en A2.
        _subtitulo = ('="Convenio Hostelería — "&TEXT(' + self.p.ref('pagas')
                      + ',"0")&" pagas + SS "&TEXT('
                      + self.p.ref('ss_empresa') + '*100,"0")&" %"')
        _hay_subtitulo = False
        for _r in range(1, cab):
            _v = ws.cell(row=_r, column=1).value
            if isinstance(_v, str) and _v.startswith('=') \
                    and 'Convenio Hostel' in _v:
                # 2.ª pasada: el subtítulo ya es la fórmula de la 1.ª
                _hay_subtitulo = True
                continue
            if not isinstance(_v, str) or _v.startswith('='):
                continue
            if not re.search(r'\d+[.,]?\d*\s*%', _v):
                continue
            if not re.search(r'convenio|pagas|seguridad social|\bss\b', _v,
                             re.I):
                continue
            self.anota('Personal!A' + str(_r) + ': subtítulo con el tipo de '
                       'Seguridad Social escrito a mano «' + _v[:60]
                       + '» → fórmula desde Supuestos (§7-bis.11)')
            fx(ws, 'A' + str(_r), _subtitulo)
            _hay_subtitulo = True
        # sólo si la fila 2 está VACÍA: el bar-restaurante trae ahí su propio
        # subtítulo («Cuadro de Personal y Coste Laboral») y pisarlo sería
        # cambiar contenido que nadie ha pedido cambiar.
        _fila2_vacia = all(ws.cell(row=2, column=_c).value in (None, '')
                           for _c in range(1, max(2, ws.max_column) + 1))
        if not _hay_subtitulo and cab > 2 and _fila2_vacia:
            fx(ws, 'A2', _subtitulo)
            self.anota('Personal!A2: la hoja no tenía subtítulo y la fila 2 '
                       'estaba vacía — se escribe por fórmula desde '
                       'Supuestos (M13 / R22-PAN-17)')
        plantilla = self.dato('PLANTILLA') or self._leer_plantilla(ws, cab)
        # el pie del fichero v1.1 lleva el tipo de Seguridad Social escrito a
        # mano («33.4%»): repetiría un parámetro que ahora vive en celda y
        # quedaría desmintiendo a la propia columna (§7-bis.11)
        pie = [t for t in _pie(ws, cab)
               if not re.search(r'\d+[.,]?\d*\s*%', t)]
        for t in _pie(ws, cab):
            if re.search(r'\d+[.,]?\d*\s*%', t):
                self.anota('Personal: fuera la nota con el porcentaje escrito '
                           'a mano «' + t[:60] + '» (§7-bis.11)')
        _limpiar_area(ws, cab, ws.max_row, 10)
        # ⚠️ Las cabeceras llevan «(€)» a propósito: `motor.formatos_por_tipo`
        # decide por la CABECERA de la columna, y «Bruto mes» contiene la
        # palabra «mes», que el motor lee como recuento y le quitaría el
        # formato de euro a toda la columna.
        cabeceras = ('Puesto', 'Personas', 'Jornada',
                     'Bruto mes (€, total del puesto)',
                     'Seg. Social a cargo de la empresa (€)',
                     'Coste mes (€, total del puesto)', 'Coste año (€)',
                     'Horas/semana por persona', 'Horas/año del puesto',
                     'Notas')
        for i, texto in enumerate(cabeceras):
            motor.val(ws, get_column_letter(i + 1) + str(cab), texto,
                      bold=True, wrap=True)
        rej = Rejilla(ws, cab + 1,
                      driver=lambda R, c, k: (R.c(k, 'D')
                                              if str(k).startswith('p_')
                                              else None),
                      cols_driver=('E', 'F', 'G', 'H', 'I'))
        self.rej['personal'] = rej
        ss = self.p.ref('ss_empresa')
        pagas = self.p.ref('pagas')
        for i, fila in enumerate(plantilla):
            (puesto, personas, bruto, nota, fuente,
             jornada) = (list(fila) + [None] * 6)[:6]
            clave = 'p_%d' % i
            rej.add(clave, rot=puesto,
                    valores={'B': personas, 'C': jornada or 1.0, 'D': bruto,
                             'J': nota or ''},
                    verdes=('B', 'C', 'D'),
                    fmt_B=motor.FMT_ENT, fmt_C=motor.FMT_PCT0,
                    fmt_D=motor.FMT_EUR, fmt_H=motor.FMT_DEC,
                    fmt_I=motor.FMT_ENT,
                    formulas={
                        'E': (lambda R, k=clave: '=' + R.c(k, 'D') + '*' + ss),
                        'F': (lambda R, k=clave: '=' + R.c(k, 'D') + '+'
                              + R.c(k, 'E')),
                        # RT-08 — un 0 en «número de pagas» ponía el coste
                        # de personal a CERO y hacía que todos los semáforos
                        # de Instrucciones dijeran CUMPLE. Nunca hay menos de
                        # doce mensualidades: por debajo se trata como 12 y
                        # la validación de datos lo impide de entrada.
                        'G': (lambda R, k=clave: '=' + R.c(k, 'F') + '*MAX(12,'
                              + pagas + ')'),
                        # RD-09 / RC-19 — la plantilla no estaba dimensionada
                        # por HORAS DE SERVICIO, que es lo que exige la
                        # decisión §7-bis.17: la hoja sólo traía «Jornada»
                        # como porcentaje, sin cuadrante ni comprobación de
                        # cobertura contra los días de apertura.
                        'H': (lambda R, k=clave: '=' + R.c(k, 'C') + '*'
                              + R.c('jornada_completa', absoluta=True)),
                        'I': (lambda R, k=clave: '=' + R.c(k, 'H') + '*'
                              + R.c(k, 'B') + '*'
                              + R.c('semanas', absoluta=True))},
                    fmt=motor.FMT_EUR)
        primero, ultimo = 'p_0', 'p_%d' % (len(plantilla) - 1)
        rej.add('total', rot='TOTAL PLANTILLA', bold=True,
                formulas=dict(
                    [(col, (lambda R, c=col:
                            '=IF(COUNT(' + R.c(primero, 'D') + ':'
                            + R.c(ultimo, 'D') + ')=0,"",SUM('
                            + R.c(primero, c) + ':' + R.c(ultimo, c) + '))'))
                     for col in ('B', 'D', 'E', 'F', 'G', 'I')]
                    + [('C', (lambda R: '=IF(COUNT(' + R.c(primero, 'D') + ':'
                              + R.c(ultimo, 'D') + ')=0,"",SUM('
                              + R.c(primero, 'C') + ':' + R.c(ultimo, 'C')
                              + '))'))]),
                fmt=motor.FMT_EUR, fmt_B=motor.FMT_ENT,
                fmt_C=motor.FMT_DEC2, fmt_I=motor.FMT_ENT)
        # RD-10 — el P&L hacía crecer el personal con el VOLUMEN, así que el
        # ratio de coste laboral salía idéntico en los tres años POR
        # CONSTRUCCIÓN y el semáforo de los años 2 y 3 no podía fallar nunca.
        # El personal es un coste ESCALONADO: se contrata gente, no un
        # porcentaje.
        # ⚠️ B7 / M-04 / REF-02-motor / MOT-01 / NUEVO-CAF-01 (2026-09-05):
        # el rótulo publicaba «, RD-10» —un id interno de auditoría— en
        # la única celda de los dos entregables que se lo enseñaba al
        # comprador. El id vive en el comentario de arriba.
        rej.add(rot='CRECIMIENTO DE LA PLANTILLA (coste escalonado)',
                bold=True)
        for anio, clave, defecto in ((2, 'alta_2', 6000), (3, 'alta_3', 6000)):
            rej.add(clave,
                    rot='Coste anual de las incorporaciones del año '
                        + str(anio) + ' (€)',
                    fmt=motor.FMT_EUR0, verdes=('B',),
                    valores={'B': defecto,
                             'J': 'Coste ANUAL completo (bruto + Seguridad '
                                  'Social × pagas) de lo que se contrate ese '
                                  'año. El P&L lo suma al coste del año '
                                  'anterior: la plantilla no crece con un '
                                  'porcentaje, crece con personas'})
        rej.add('coste_2', rot='Coste de personal del año 2',
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=IF(' + R.c('total', 'G')
                                + '="","",' + R.c('total', 'G') + '*(1+'
                                + self.p.ref('ipc') + ')+' + R.c('alta_2')
                                + ')')})
        rej.add('coste_3', rot='Coste de personal del año 3',
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=IF(' + R.c('coste_2')
                                + '="","",' + R.c('coste_2') + '*(1+'
                                + self.p.ref('ipc') + ')+' + R.c('alta_3')
                                + ')')})
        # ---- cobertura por horas de servicio (RD-09 / RC-19) -------------
        rej.add(rot='COBERTURA POR HORAS DE SERVICIO', bold=True)
        rej.add('jornada_completa',
                rot='Jornada completa del convenio (horas/semana)',
                fmt=motor.FMT_ENT, verdes=('B',),
                valores={'B': 40,
                         'J': 'Art. 34.1 del Estatuto de los Trabajadores; tu '
                              'convenio provincial puede fijar menos'})
        # ⚠️ M16 / R22-PAN-21 (2026-09-05). La nota decía «52 menos las
        # vacaciones del art. 38 ET (30 días naturales)», que son 4,3
        # semanas: 47,7, no 46. El número (conservador, y es el denominador
        # de toda la cobertura de horas) se queda; la nota pasa a decir lo
        # que de verdad descuenta.
        _SEMANAS = 46
        rej.add('semanas', rot='Semanas trabajadas al año por persona',
                fmt=motor.FMT_ENT, verdes=('B',),
                valores={'B': _SEMANAS,
                         'J': str(_SEMANAS) + ' semanas: 52 menos las '
                              'vacaciones (30 días naturales, art. 38 ET), '
                              'los festivos y las ausencias medias'})
        # ⚠️ A4 / MOT-03 / REF-06 / M-02 (2026-09-05). Estas dos celdas son el
        # NUMERADOR del semáforo de cobertura y, con él, quien dimensiona la
        # plantilla entera y el labour cost. Estaban CABLEADAS al cuadrante de
        # un restaurante (13 h × 2 personas) y no tenían clave de contenido:
        # el food truck salía al 45,6 % (rojo) y la panadería en verde por
        # 17 horas al año dejando fuera el turno de obrador que el propio
        # libro staffea. Ahora el molde declara su cuadrante, con los valores
        # de siempre por defecto, y admite una SEGUNDA franja (producción
        # anterior a la apertura) para los formatos que la tengan.
        cob = self.dato('COBERTURA', {}) or {}
        h_prod = _num(cob.get('horas_produccion'), 0) or 0
        p_prod = _num(cob.get('personas_produccion'), 0) or 0
        rej.add('horas_dia', rot='Horas de servicio al día que hay que cubrir',
                fmt=motor.FMT_DEC, verdes=('B',),
                valores={'B': _num(cob.get('horas_dia'), 13),
                         'J': cob.get('nota_horas')
                         or 'De la apertura al cierre, barra incluida. '
                            'Cuéntalas sobre tu horario real'})
        rej.add('personas_franja',
                rot='Personas necesarias a la vez en cada franja',
                fmt=motor.FMT_DEC, verdes=('B',),
                valores={'B': _num(cob.get('personas_franja'), 2),
                         'J': cob.get('nota_personas')
                         or 'Media de presencia simultánea entre cocina, '
                            'barra y sala. En los dos servicios fuertes '
                            'hará falta más y a media tarde menos'})
        if h_prod and p_prod:
            rej.add('horas_produccion',
                    rot='Horas de producción al día ANTES de abrir',
                    fmt=motor.FMT_DEC, verdes=('B',),
                    valores={'B': h_prod,
                             'J': cob.get('nota_horas_produccion')
                             or 'Turno anterior a la apertura (obrador, '
                                'mise en place): no se solapa con el horario '
                                'de atención al público'})
            rej.add('personas_produccion',
                    rot='Personas en el turno de producción',
                    fmt=motor.FMT_DEC, verdes=('B',),
                    valores={'B': p_prod,
                             'J': cob.get('nota_personas_produccion')
                             or 'Presencia simultánea en ese turno'})
        # B1 / REF-15 — era la única fórmula del bloque sin la guarda de
        # vacío que llevan sus vecinas: con el libro vaciado imprimía un «0»
        # suelto en vez de dejar la celda en blanco.
        _prod = ('+' + '{hp}' + '*' + '{pp}' + '*'
                 + self.p.ref('dias_apertura')) if (h_prod and p_prod) else ''
        rej.add('horas_necesarias', rot='Horas de servicio a cubrir al año',
                fmt=motor.FMT_ENT,
                formulas={'B': (lambda R: '=IF(OR(' + R.c('horas_dia')
                                + '="",' + R.c('personas_franja') + '="",'
                                + self.p.ref('dias_apertura') + '=""),"",'
                                + R.c('horas_dia') + '*'
                                + R.c('personas_franja') + '*'
                                + self.p.ref('dias_apertura')
                                + _prod.replace(
                                    '{hp}', R.c('horas_produccion')
                                    if _prod else '')
                                       .replace('{pp}',
                                                R.c('personas_produccion')
                                                if _prod else '')
                                + ')')})
        rej.add('horas_contratadas', rot='Horas contratadas al año',
                fmt=motor.FMT_ENT,
                formulas={'B': (lambda R: '=' + R.c('total', 'I'))})
        rej.add('cobertura', rot='Cobertura (contratadas / necesarias)',
                fmt=motor.FMT_PCT, bold=True,
                formulas={'B': (lambda R: '=' + R.c('horas_contratadas') + '/'
                                + R.c('horas_necesarias')),
                          'J': '="Por debajo del 100 % la plantilla no llega '
                               'a cubrir el horario que declara el plan: o se '
                               'contrata más o se recorta el horario"'})
        motor.semaforo_num(ws, rej.c('cobertura') + ':' + rej.c('cobertura'),
                           verde_si=rej.c('cobertura') + '>=1',
                           ambar_si=rej.c('cobertura') + '>=0.9',
                           rojo_si=rej.c('cobertura') + '<0.9')
        self.pendientes.append(rej)
        # El bruto por PERSONA nunca por debajo del suelo legal (§2.6) — y en
        # las jornadas parciales, de ese suelo en PROPORCIÓN a la jornada.
        # RD-25: el suelo real de la hostelería es el CONVENIO PROVINCIAL, no
        # el SMI, y el propio libro lo declara aplicable dos filas más abajo.
        # El semáforo compara contra el MAYOR de los dos; con la celda de
        # convenio a 0 (que es como se entrega, porque la tabla salarial es
        # provincial y no se inventa) compara sólo contra el SMI.
        rango = rej.c(primero, 'D') + ':' + rej.c(ultimo, 'D')
        smi = self.p.ref('smi_anual')
        convenio = self.p.ref('salario_convenio')
        base = ('IFERROR(' + rej.c(primero, 'D') + '/' + rej.c(primero, 'B')
                + ',0)*' + pagas)
        motor.semaforo_num(ws, rango,
                           rojo_si=base + '<MAX(' + smi + ',' + convenio
                           + ')*' + rej.c(primero, 'C'))
        fila_nota = rej.ultima + 2
        motor.val(ws, 'A' + str(fila_nota),
                  'Las columnas «Bruto mes», «Seg. Social», «Coste mes» y '
                  '«Coste año» son TOTALES de la fila: en un puesto con dos '
                  'personas incluyen a las dos. «Jornada» es el porcentaje '
                  'sobre la jornada completa del convenio y sirve para '
                  'comparar el salario con el suelo legal en proporción. El '
                  'tipo de Seguridad Social y el número de pagas están en la '
                  'hoja «0. Supuestos».', wrap=True)
        motor.val(ws, 'A' + str(fila_nota + 1),
                  'El suelo salarial es el MAYOR entre el SMI y el salario de '
                  'convenio, los dos en la hoja «0. Supuestos»; las jornadas '
                  'parciales lo llevan en proporción. El convenio aplicable '
                  'es el PROVINCIAL de hostelería: no existe una tabla '
                  'salarial estatal única, así que la celda del convenio se '
                  'entrega vacía para que copies la de tu provincia.',
                  wrap=True)
        for i, texto in enumerate(pie):
            motor.val(ws, 'A' + str(fila_nota + 2 + i), texto)
        motor.anchos(ws, {'A': 34, 'B': 10, 'C': 10, 'D': 16, 'E': 18,
                          'F': 16, 'G': 16, 'H': 14, 'I': 14, 'J': 40})
        motor.print_setup(ws, header_row=cab)
        return rej

    def _leer_plantilla(self, ws, cab):
        """Plantilla del fichero, homogeneizada a totales de fila (TEC-16)."""
        cabeceras = dict((motor.norm(c.value), c.column) for c in ws[cab]
                         if isinstance(c.value, str))
        col_personas = cabeceras.get('personas')
        # RT-02 — la lista cerrada no reconocía la cabecera que escribe este
        # mismo grupo («Bruto mes (€, total del puesto)»), así que en la 2.ª
        # pasada `col_bruto` caía al índice 3, que es la columna «Jornada»,
        # y los siete sueldos de la hoja Personal se reescribían a 1 €. Con
        # ejecución real en serie y con reintentos, eso es un fichero de
        # producción destruido: la columna se busca por RÓTULO NORMALIZADO.
        col_bruto = None
        for clave, col in sorted(cabeceras.items(), key=lambda kv: kv[1]):
            if clave.startswith('bruto') or clave.startswith('salario bruto'):
                col_bruto = col
                break
        if col_bruto is None:
            self.anota('Personal: no se reconoce la columna de bruto en '
                       + ws.title + ' (cabeceras: '
                       + ', '.join(sorted(cabeceras)) + '): se lee la '
                       'plantilla del módulo de contenido, no del fichero')
            return []
        col_nota = cabeceras.get('notas')
        fuera = []
        for r in range(cab + 1, ws.max_row + 1):
            rot = motor._rotulo_de_fila(ws, r, max_col=1)
            if not rot or motor.RX_TOTAL.match(rot) \
                    or re.match(r'^(\*|NOTA|ChefBusiness)', rot):
                continue
            bruto = _num(ws.cell(row=r, column=col_bruto).value)
            if bruto is None:
                continue
            personas = _num(ws.cell(row=r, column=col_personas).value, 1) \
                if col_personas else 1
            nota = ws.cell(row=r, column=col_nota).value if col_nota else None
            fuera.append((rot, int(personas), bruto * personas,
                          nota if isinstance(nota, str) else None,
                          'leído del fichero (bruto × personas)', 1.0))
        return fuera

    # -- §2.2 -------------------------------------------------------------
    def inversion(self):
        """`Inversión Inicial`: subtotales, fondo calculado e IVA (§2.2)."""
        ws = self.ws_inversion
        cab = _cabecera(ws)
        canon = self.dato('INVERSION_CANON') or RX_CANON_INV
        partidas = _partidas(ws, cab, canon=canon)
        reglas = self._reglas('INVERSION')
        extras = self.dato('INVERSION_EXTRA', []) or []
        tablas = self.dato('AMORTIZABLE') or AMORT_DEFECTO
        pie = _pie(ws, cab)
        _limpiar_area(ws, cab, ws.max_row, 5)
        # RD-14 / RT-25 — el IVA soportado se calculaba al 21 % sobre TODA la
        # inversión menos fianza y fondo, incluyendo tasas municipales (no
        # sujetas) y seguros (exentos, art. 20.Uno.16 LIVA). Ahora cada
        # partida declara si lleva IVA en una columna con desplegable y B46
        # es un SUMIF: el comprador puede corregir la clasificación de su
        # caso sin tocar una fórmula.
        for i, texto in enumerate(('Concepto', 'Importe', '% s/inversión',
                                   'Notas', '¿Lleva IVA?')):
            motor.val(ws, get_column_letter(i + 1) + str(cab), texto,
                      bold=True)
        # agrupar por bloque conservando el orden
        bloques, orden = {}, []
        for _r, rot, importe, nota, bloque in partidas:
            # el rótulo se limpia ANTES de consultar las reglas: en la 2.ª
            # pasada ya viene limpio y la clave tiene que ser la misma o el
            # fichero dejaría de ser idempotente
            rot, cambiado = _limpiar_rotulo(rot)
            if cambiado:
                self.anota('Inversión: rótulo con parámetro escrito a mano → '
                           '«' + rot + '» (§7-bis.11)')
            accion = reglas.get(motor.norm(rot))
            if accion and accion[0] == 'suprimir':
                self.anota('Inversión: fuera «' + rot + '» — '
                           + (accion[1] or 'lo pide el módulo de contenido'))
                continue
            if accion and motor._es_numero(accion[0]):
                self.anota('Inversión: «' + rot + '» ' + str(importe) + ' → '
                           + str(accion[0]) + ' — ' + (accion[1] or ''))
                importe, nota = float(accion[0]), accion[1] or nota
            elif accion and accion[0] is None and accion[1]:
                # importe None = «deja la cifra del fichero, cambia la nota»
                nota = accion[1]
            # A6 / M-09 — el rótulo de una fila PRESERVADA también se puede
            # corregir sin moverla de sitio (antes había que suprimirla y
            # darla de alta por INVERSION_EXTRA, lo que mandaba la partida
            # más cara del plan al final del bloque).
            if accion and len(accion) > 2 and accion[2] \
                    and motor.norm(accion[2]) != motor.norm(rot):
                self.anota('Inversión: rótulo «' + rot + '» → «' + accion[2]
                           + '» (A6)')
                rot = accion[2]
            bloque = bloque or 'INVERSIÓN'
            if bloque not in bloques:
                bloques[bloque] = []
                orden.append(bloque)
            bloques[bloque].append((rot, importe, nota))
        # las partidas nuevas se añaden UNA vez: en la 2.ª pasada ya están
        # dentro de su bloque y `_partidas` las devuelve como preservadas.
        # Sin esta guarda, la segunda pasada duplicaba la terraza y las dos
        # líneas de existencias y desplazaba la hoja entera (idempotencia
        # rota, que es la red de seguridad que exige la propia SPEC).
        ya = set(motor.norm(r) for b in bloques for r, _i, _n in bloques[b])
        for bloque, rot, importe, nota, _fuente in [
                tuple(list(e) + [None] * 5)[:5] for e in extras]:
            if motor.norm(rot) in ya:
                continue
            ya.add(motor.norm(rot))
            bloque = bloque or (orden[0] if orden else 'INVERSIÓN')
            if bloque not in bloques:
                bloques[bloque] = []
                orden.append(bloque)
            bloques[bloque].append((rot, importe, nota))

        rej = Rejilla(ws, cab + 1)
        self.rej['inversion'] = rej
        alq = self.p.ref('alquiler_mes')
        fianza_m = self.p.ref('fianza_meses')
        meses_previos = self.p.ref('meses_renta_previa')
        pct_imp = self.p.ref('pct_imprevistos')
        claves_bloque, amortiza = [], {'obra': [], 'maquinaria': []}
        con_iva = []
        #: B5 / R-18 — las filas CALCULADAS (fianza, renta previa,
        #: imprevistos, colchón) también entran en el SUMIF del IVA, pero su
        #: celda de marca está BLOQUEADA: colgarles el desplegable enseña una
        #: lista que Excel se niega a aplicar. Sólo las partidas de input.
        editables_iva = []
        #: A8 — rótulo de cada fila de la inversión, en ORDEN, para generar
        #: desde el libro la enumeración de lo que NO se amortiza (la lista
        #: estaba escrita a mano en el generador y nombraba partidas que este
        #: producto puede no tener: REF-19).
        rotulos_inv = []
        # un bloque sin partidas (el «FONDO DE MANIOBRA» original, cuya única
        # línea es canónica y la regenera este grupo) no se escribe: dejarlo
        # produciría un `SUM` sobre sí mismo, que Excel marca como referencia
        # circular
        orden = [b for b in orden if bloques.get(b)]
        ultimo_bloque = None
        for bloque in orden:
            clave_b = 'b_' + str(len(claves_bloque))
            claves_bloque.append(clave_b)
            rej.add(clave_b, rot=bloque, bold=True, fmt=motor.FMT_EUR0,
                    fmt_C=motor.FMT_PCT)
            primero = None
            ultimo_item = None
            for rot, importe, nota in bloques[bloque]:
                clave = 'i_%d' % len(rej.filas)
                primero = primero or clave
                ultimo_item = clave
                grupo = ('no' if RX_BLOQUE_NO_AMORT.search(bloque or '')
                         else _clasificar_amortizable(rot, tablas))
                if grupo in amortiza:
                    amortiza[grupo].append(clave)
                con_iva.append(clave)
                editables_iva.append(clave)
                rotulos_inv.append((clave, rot))
                rej.add(clave, rot=rot,
                        valores={'B': importe, 'D': nota or '',
                                 'E': _lleva_iva(rot)},
                        # RT-12/RC-16: verde SÓLO lo que alimenta una fórmula
                        # (el importe y el marcador de IVA). Pintar de verde
                        # la columna de notas convierte el comentario del
                        # autor en un campo de entrada y contradice a
                        # Instrucciones: cambiar una nota no recalcula nada.
                        fmt=motor.FMT_EUR0, verdes=('B', 'E'),
                        formulas={'C': (lambda R, k=clave:
                                        '=' + R.c(k) + '/'
                                        + R.c('total', absoluta=True))},
                        fmt_C=motor.FMT_PCT)
            # las derivadas del alquiler y el colchón de obra, en su bloque
            if bloque == orden[0]:
                rej.add('fianza', rot='Fianza del alquiler',
                        fmt=motor.FMT_EUR0, fmt_C=motor.FMT_PCT,
                        valores={'E': 'No'},
                        formulas={
                            'B': ('=IF(' + alq + '="","",' + alq + '*'
                                  + fianza_m + ')'),
                            'C': (lambda R: '=' + R.c('fianza') + '/'
                                  + R.c('total', absoluta=True)),
                            'D': texto_num(fianza_m, '', ' meses de renta, al '
                                           'tipo de la hoja de Supuestos. Es '
                                           'un activo recuperable al cerrar: '
                                           'ni se amortiza ni lleva IVA')})
                # RT-14 / RD-01 — «Primera mensualidad de alquiler» y el mes 1
                # del P&L eran EL MISMO euro contado dos veces. La renta que
                # es inversión es la de los meses ANTERIORES a la apertura
                # (obra y licencias), que el P&L no cubre porque arranca el
                # día que se abre.
                rej.add('renta1',
                        rot='Alquiler de los meses previos a la apertura',
                        fmt=motor.FMT_EUR0, fmt_C=motor.FMT_PCT,
                        valores={'E': 'Sí'},
                        formulas={
                            'B': ('=IF(' + alq + '="","",' + alq + '*'
                                  + meses_previos + ')'),
                            'C': (lambda R: '=' + R.c('renta1') + '/'
                                  + R.c('total', absoluta=True)),
                            # B8 / PLURAL-01 / NUEVO-CAF-02 / REF-23: con
                            # `meses_renta_previa` = 1, que es el valor por
                            # defecto del grupo A, la nota imprimía «Renta de
                            # los 1 meses». El plural se compone por fórmula.
                            'D': ('="Renta "&IF(' + meses_previos
                                  + '=1,"del mes","de los "&TEXT('
                                  + meses_previos + ',"0")&" meses")&" de '
                                  'obra y licencias. NO se solapa con el '
                                  'P&L, que arranca el día de la apertura: '
                                  'los doce meses del P&L son los doce de '
                                  'explotación"')})
                # RD-02 / RC-02 — imprevistos de obra. Por FÓRMULA sobre las
                # partidas del bloque, nunca literal (§7-bis.11), y sin
                # incluirse a sí mismo ni a las derivadas: un `SUM` sobre el
                # subtotal del bloque sería referencia circular.
                if primero and ultimo_item:
                    # ⚠️ M4 / R22-CAF-11 / R22-PAN-20 / REF22-BAR-06
                    # (2026-09-05). La enumeración de la nota era una LISTA
                    # FIJA («proyecto técnico, obra civil, instalaciones,
                    # extracción, decoración y rotulación») heredada del
                    # molde del bar: en la cafetería nombraba una
                    # «extracción» que no existe (licencia inocua, sin
                    # campana) y se dejaba fuera la fontanería, que sí entra;
                    # en la panadería nombraba una «decoración» inexistente.
                    # Se compone desde los RÓTULOS de las filas que de
                    # verdad entran en la fórmula, igual que ya hace
                    # `_enumerar_fuera()` con las no amortizables.
                    # ⚠️ A8 / REF-06 (2026-09-05). La base era `SUM` del
                    # BLOQUE ENTERO, así que el 10 % de «colchón de obra» se
                    # dotaba también sobre la comisión de la inmobiliaria, la
                    # licencia de actividad, el marketing de lanzamiento, el
                    # stock inicial y el TPV — 10.500 € que no son obra en
                    # tapas-bar—, mientras la nota que el propio motor
                    # generaba afirmaba «de las partidas de obra y
                    # acondicionamiento de este bloque». Ahora la base son
                    # exactamente las partidas que `AMORT_DEFECTO` clasifica
                    # como 'obra' en ESTE bloque, y la nota dice la verdad.
                    obra_bloque = [k for k in amortiza['obra']
                                   if rej.fila(k) >= rej.fila(primero)]
                    rej.add('imprevistos',
                            rot='Imprevistos de obra y acondicionamiento',
                            fmt=motor.FMT_EUR0, fmt_C=motor.FMT_PCT,
                            valores={'E': 'Sí'},
                            formulas={
                                'B': ((lambda R, ks=tuple(obra_bloque):
                                       '=IF(COUNT('
                                       + ','.join(R.c(k) for k in ks)
                                       + ')=0,"",ROUND(('
                                       + '+'.join(R.c(k) for k in ks) + ')*'
                                       + pct_imp + ',0))')
                                      if obra_bloque else
                                      (lambda R, a=primero, b=ultimo_item:
                                       '=IF(COUNT(' + R.c(a) + ':' + R.c(b)
                                       + ')=0,"",ROUND(SUM(' + R.c(a) + ':'
                                       + R.c(b) + ')*' + pct_imp + ',0))')),
                                'C': (lambda R: '=' + R.c('imprevistos') + '/'
                                      + R.c('total', absoluta=True)),
                                'D': texto_pct(
                                    pct_imp, 'Al ', ' de las partidas de OBRA '
                                    'Y ACONDICIONAMIENTO de este bloque ('
                                    + _enumerar_base(
                                        rotulos_inv, obra_bloque, primero,
                                        ultimo_item, rej)
                                    + '), no del bloque entero: el resto de '
                                    'partidas del bloque quedan fuera. Un '
                                    'banco no financia una reforma sin '
                                    'colchón: el porcentaje está en la hoja '
                                    'de Supuestos')})
                    con_iva.append('imprevistos')
                    rotulos_inv.append(('imprevistos',
                                        'Imprevistos de obra y '
                                        'acondicionamiento'))
                    # A8 / REF-19 — el colchón de obra es MAYOR COSTE DE LA
                    # OBRA: capitaliza y entra en la base amortizable. Sin
                    # esto, el libro amortizaba una obra más barata que la
                    # que financia. No hay circularidad: la fila cuelga de
                    # las partidas de obra, que son inputs directos.
                    if obra_bloque:
                        amortiza['obra'].append('imprevistos')
                con_iva.append('renta1')
                rotulos_inv.append(('fianza', 'Fianza del alquiler'))
                rotulos_inv.append(('renta1',
                                    'Alquiler de los meses previos a la '
                                    'apertura'))
            ultimo_b = rej.filas[-1]['clave']
            # A7 / REF-14 — el SUMIF del IVA soportado terminaba en la última
            # clave de `con_iva`, que en el molde A-β (lista plana, un solo
            # bloque) es «renta1»: la fila CALCULADA de imprevistos, marcada
            # «Sí» y sumada por el subtotal (`SUM(B7:B27)`), quedaba FUERA del
            # SUMIF (`$E$7:$E$26`). En tapas-bar eran 9.500 € × 21 % = 1.995 €
            # de IVA soportado sin contar, y con ellos la necesidad de caja.
            # El rango del SUMIF pasa a acabar donde acaba el SUM del último
            # subtotal, que es esta misma fila.
            ultimo_bloque = ultimo_b
            rej.filas[rej.fila(clave_b) - rej.fila0]['formulas'] = {
                # RT-18 — `SUM` sobre celdas vacías devuelve 0, no `""`:
                # un libro en blanco imprimía «0 €» en los seis subtotales
                'B': ((lambda R, a=primero, b=ultimo_b:
                       '=IF(COUNT(' + R.c(a) + ':' + R.c(b) + ')=0,"",SUM('
                       + R.c(a) + ':' + R.c(b) + '))')
                      if primero and ultimo_b != clave_b else '=0'),
                'C': (lambda R, k=clave_b: '=' + R.c(k) + '/'
                      + R.c('total', absoluta=True))}
        # fondo de maniobra (TEC-07, DOM-12, NUEVO-01)
        clave_b = 'b_fondo'
        claves_bloque.append(clave_b)
        rej.add(clave_b, rot='FONDO DE MANIOBRA', bold=True,
                fmt=motor.FMT_EUR0, fmt_C=motor.FMT_PCT,
                formulas={'B': (lambda R: '=' + R.c('fondo')),
                          'C': (lambda R: '=' + R.c(clave_b) + '/'
                                + R.c('total', absoluta=True))})
        meses = self.p.ref('meses_fondo')
        # RD-16 / RC-18 / RT-26 — el colchón se dotaba sobre una magnitud
        # CONTABLE: los tres meses salían de unos costes fijos que incluyen la
        # amortización, que NO es salida de caja. Se financiaban con deuda a
        # siete años 2.220 € de un gasto que nadie paga. Los intereses sí son
        # salida de caja y se quedan dentro.
        caja_fija = ('(' + self.rej['pyg'].r('tcf') + '-'
                     + self.rej['pyg'].r('cf_amort') + ')')
        rotulos_inv.append(('fondo',
                            'Colchón operativo hasta alcanzar el equilibrio'))
        rej.add('fondo', rot='Colchón operativo hasta alcanzar el equilibrio',
                fmt=motor.FMT_EUR0, fmt_C=motor.FMT_PCT,
                valores={'E': 'No'},
                formulas={
                    # ⚠️ una referencia a una celda VACÍA devuelve 0, no
                    # `""`: sin este guarda, un libro en blanco dota un fondo
                    # de «0 €» y el semáforo de la tesorería se pone verde
                    'B': (lambda R: '=IF(' + meses + '*' + caja_fija
                          + '=0,"",' + meses + '*' + caja_fija + '/12)'),
                    'C': (lambda R: '=' + R.c('fondo') + '/'
                          + R.c('total', absoluta=True)),
                    # B8 (misma familia que la renta previa): con
                    # `meses_fondo` = 1 la nota decía «Cubre 1 meses». Con la
                    # concordancia por fórmula, el caso base (3) no cambia.
                    'D': ('="Cubre "&TEXT(' + meses + ',"0")&IF(' + meses
                          + '=1," mes"," meses")&" de costes fijos DE CAJA '
                          'del año 1 (los del P&L menos la amortización, que '
                          'no se paga), que es el mínimo que exigen las '
                          'Instrucciones de este libro"')})
        rej.add('total', rot='INVERSIÓN TOTAL (suma de los bloques)',
                bold=True, fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=' + '+'.join(
                    R.c(k) for k in claves_bloque)),
                    # RC-33 — la columna de % divide entre el total CON
                    # colchón, así que los bloques de activo se leen más
                    # pequeños de lo que son. Se dice en la propia hoja y se
                    # publica el capex aparte.
                    'D': (lambda R: '="Los porcentajes de la columna C se '
                          'leen sobre esta cifra, que incluye el fondo de '
                          # la coma de «0,0%» NO es un decimal: en un
                          # código de formato es el separador de MILES, así
                          # que ya imprimía «34%» y no «33,9%». Se escribe
                          # «0%», que es lo que de verdad hace.
                          'maniobra ("&TEXT(' + R.c('b_fondo', 'C')
                          + ',"0%")&" del total). Para comparar el '
                          'desglose con otro plan, usa la fila de CAPEX de '
                          'abajo."')})
        rej.add('capex', rot='CAPEX (inversión sin el fondo de maniobra)',
                bold=True, fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=' + R.c('total') + '-'
                                + R.c('b_fondo')),
                          'D': '="Es lo que se compra: obra, equipamiento, '
                               'lanzamiento y gastos de constitución. El '
                               'colchón de caja no es inversión, es '
                               'circulante"'})
        # IVA soportado (TEC-11, RD-14, RT-25): sólo sobre lo que lo lleva
        rej.add('iva', rot='IVA soportado sobre la inversión (recuperable)',
                fmt=motor.FMT_EUR0,
                formulas={
                    'B': (lambda R, ks=tuple(con_iva),
                          fin=(ultimo_bloque or (con_iva[-1] if con_iva
                                                 else None)):
                          ('=IF(' + R.c('total') + '="","",SUMIF('
                           + R.c(ks[0], 'E', absoluta=True) + ':'
                           + R.c(fin, 'E', absoluta=True) + ',"Sí",'
                           + R.c(ks[0], absoluta=True) + ':'
                           + R.c(fin, absoluta=True) + ')*'
                           + self.p.ref('iva_soportado') + ')') if ks
                          else '=0'),
                    # M15 / R22-PAN-19 — la nota justificaba con «las
                    # tasas y licencias no están sujetas» una fila
                    # («Constitución SL + notaría») que mezcla la tasa
                    # registral con el ARANCEL NOTARIAL, que es un servicio
                    # profesional sujeto y no exento. La clasificación de la
                    # fila no se toca (la decide el comprador en su
                    # desplegable); lo que se corrige es la afirmación
                    # fiscal.
                    'D': texto_pct(self.p.ref('iva_soportado'),
                                   'Al ', ' sobre las partidas marcadas con '
                                   '«Sí» en la columna de IVA. Las tasas y '
                                   'licencias no están sujetas y los seguros '
                                   'están exentos (art. 20.Uno.16 LIVA). Las '
                                   'tasas y aranceles registrales no llevan '
                                   'IVA; los honorarios de notaría y '
                                   'gestoría sí (21 %, deducible). Se '
                                   'recupera con el modelo 303, pero hay que '
                                   'adelantarlo')})
        rej.add('caja', rot='NECESIDAD TOTAL DE CAJA AL ARRANQUE', bold=True,
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=' + R.c('total') + '+'
                                + R.c('iva')),
                          'D': '="Es la cifra que tiene que cubrir la hoja de '
                               'Financiación"'})
        # bases de amortización (TEC-20, NUEVO-02)
        def _enumerar_fuera():
            dentro = set(amortiza['obra']) | set(amortiza['maquinaria'])
            fuera_ = [r.replace('"', '').strip()
                      for k, r in rotulos_inv if k not in dentro]
            # sin duplicados y conservando el orden de la hoja
            vistos, limpio = set(), []
            for r in fuera_:
                if motor.norm(r) in vistos:
                    continue
                vistos.add(motor.norm(r))
                limpio.append(r)
            if not limpio:
                return 'ninguna partida de este libro'
            if len(limpio) == 1:
                return limpio[0]
            return ', '.join(limpio[:-1]) + ' ni ' + limpio[-1]

        rej.add('b_amort', rot='BASES DE AMORTIZACIÓN (no suman a la '
                'inversión)', bold=True)
        for grupo, etiqueta, vida in (
                ('obra', 'Obra, instalaciones y acondicionamiento',
                 self.p.ref('vida_obra')),
                ('maquinaria', 'Maquinaria, mobiliario y equipos',
                 self.p.ref('vida_maquinaria'))):
            claves = amortiza[grupo]
            rej.add('base_' + grupo, rot=etiqueta, fmt=motor.FMT_EUR0,
                    formulas={
                        # ⚠️ la guarda NO puede colgar de la inversión
                        # total: `total` incluye el fondo de maniobra, que
                        # sale de los costes fijos del P&L, que incluyen la
                        # amortización, que sale de ESTA base. Sería
                        # referencia circular. Se cuelga de la primera
                        # partida amortizable, que es un input directo.
                        'B': (lambda R, ks=tuple(claves):
                              ('=IF(' + R.c(ks[0]) + '="","",'
                               + '+'.join(R.c(k) for k in ks) + ')') if ks
                              else '=0'),
                        'D': texto_num(vida, 'Se amortiza en ', ' años')})
        rej.add('amort', rot='Amortización anual del inmovilizado', bold=True,
                fmt=motor.FMT_EUR0,
                formulas={
                    # RT-08 — con la vida útil a 0 (valor que la
                    # validación admitía) la división quedaba en `""`, `SUM`
                    # la ignoraba y el resultado antes de impuestos SUBÍA
                    # 8.882,50 € sin un solo aviso: la guarda no protegía,
                    # escondía. Con `MAX(1;…)` un cero se amortiza en un año
                    # —el resultado empeora muchísimo y se ve—, nunca mejora.
                    'B': (lambda R: '=' + R.c('base_obra') + '/MAX(1,'
                          + self.p.ref('vida_obra') + ')+'
                          + R.c('base_maquinaria') + '/MAX(1,'
                          + self.p.ref('vida_maquinaria') + ')'),
                    # RT-15 — la nota describía un libro que no era éste:
                    # nombraba stock e imprevistos, que entonces no existían
                    # como fila. Ahora se GENERA con la cifra realmente
                    # excluida, así que no puede volver a desfasarse.
                    # A8 / REF-19 — y la ENUMERACIÓN también se genera, desde
                    # las filas que de verdad quedan fuera de las dos bases:
                    # la lista escrita a mano no citaba ni la licencia de
                    # actividad ni los imprevistos, que sí estaban dentro de
                    # los euros que la propia frase declaraba excluidos.
                    'D': (lambda R: '="No se amortizan, por no ser '
                          'inmovilizado, ' + _enumerar_fuera() + '. Suman '
                          # TEC-21 residual (decisión del orquestador,
                          # 2026-08-29): «#,##0» es un formato con separador
                          # de MILES, y el separador de un código de formato
                          # se guarda en canónico anglosajón: la caché salía
                          # «Suman 95,205 € de los 179,165 €» y Excel, en
                          # español, escribe «95.205». Con «0» las dos
                          # coinciden exactamente.
                          '"&TEXT(' + R.c('total') + '-' + R.c('base_obra')
                          + '-' + R.c('base_maquinaria')
                          + ',"0")&" € de los "&TEXT(' + R.c('total')
                          + ',"0")&" € de inversión."')})
        self.pendientes.append(rej)
        motor.semaforo_num(ws, rej.c('caja') + ':' + rej.c('caja'),
                           verde_si=rej.c('caja') + '>0')
        # el marcador de IVA es una lista cerrada: si el usuario teclea «si»
        # sin tilde, el SUMIF deja de verlo y el IVA baja sin avisar (§1.5)
        if con_iva:
            filas_dv = sorted(rej.fila(k) for k in editables_iva)
            tramos, ini, prev = [], None, None
            for _f in filas_dv:
                if ini is None:
                    ini = prev = _f
                elif _f == prev + 1:
                    prev = _f
                else:
                    tramos.append((ini, prev))
                    ini = prev = _f
            if ini is not None:
                tramos.append((ini, prev))
            celdas_dv = [('E%d:E%d' % t) if t[0] != t[1] else ('E%d' % t[0])
                         for t in tramos]
            motor.dv_lista(ws, rej.c(con_iva[0], 'E') + ':'
                           + rej.c('fondo', 'E'), ['Sí', 'No'],
                           titulo='Marca no válida',
                           mensaje='Escribe «Sí» o «No»: de esta columna sale '
                                   'el IVA soportado de la inversión.',
                           celdas=celdas_dv or None)
        fila = rej.ultima + 2
        for i, texto in enumerate(pie):
            motor.val(ws, 'A' + str(fila + i), texto, wrap=True)
        motor.anchos(ws, {'A': 46, 'B': 15, 'C': 13, 'D': 58, 'E': 12})
        motor.print_setup(ws, header_row=cab)
        return rej

    # -- §2.3 -------------------------------------------------------------
    def pyg(self):
        """`P&L 3 Años`: la cadena completa desde los supuestos (§2.3)."""
        ws = self.ws_pyg
        cab = _cabecera(ws)
        fijos = self._fijos(ws, cab)
        pie = _pie(ws, cab)
        # A3 — la columna H la usa el molde que necesita partir una línea de
        # venta entre dos tipos de IVA (el pan común al 4 %); se limpia
        # SIEMPRE aunque no se escriba, o la 2.ª pasada de un producto que
        # deje de usarla se quedaría con el resto de la primera.
        _limpiar_area(ws, cab, ws.max_row, 8)
        # A3 / REF-04 — la columna ya no es sólo del IVA SOPORTADO: las
        # líneas de venta llevan en ella el tipo que REPERCUTEN, igual que
        # las de coste llevan el que soportan.
        for i, texto in enumerate(('Concepto', 'Año 1', 'Año 2', 'Año 3',
                                   '% s/ventas (año 1)', 'Notas',
                                   'Tipo de IVA de la línea (%)')):
            motor.val(ws, get_column_letter(i + 1) + str(cab), texto,
                      bold=True, wrap=True)
        # ⚠️ la columna G (tipo de IVA) NO tiene columna gemela en la fila de
        # ingresos: su driver es el del AÑO 1, o el `IF(G10="","",…)` dejaría
        # todos los tipos en blanco y el IVA soportado del libro en cero.
        rej = Rejilla(ws, cab + 1,
                      driver=lambda R, c, k: R.c('ingresos',
                                                 'B' if c == 'G' else c),
                      cols_driver=('B', 'C', 'D', 'G'))
        self.rej['pyg'] = rej
        P = self.p.ref
        crec2, crec3, ipc = P('crec_a2'), P('crec_a3'), P('ipc')
        # RD-12 — el 21 % se repercutía a TODA la línea de bebida, que el
        # propio rótulo define como «cañas, vinos, cafés y copas». Los cafés,
        # refrescos, aguas y zumos van al tipo reducido de hostelería. El
        # peso del alcohol está en celda.
        # RD-17 (2026-09-05) — y la parte alcohólica servida en sala TAMPOCO
        # se REPERCUTE al general (art. 91.Uno.2.2 de la Ley del IVA): las
        # ventas leen su propio parámetro `iva_bebida` (B63, 10 %), nunca
        # `iva_general`. Dos mezclas distintas, porque el IVA tiene dos
        # sentidos: `iva_bebida` es lo que se repercute al cliente en sala;
        # `iva_bebida_compra` es lo que se SOPORTA al comprar al proveedor,
        # que factura el alcohol al tipo general (entrega de bienes).
        # El alcohol que sale por delivery es una entrega de bienes y va al
        # general: el canal ya está en celda (`pct_delivery`, B16, 0 por
        # defecto), así que la mezcla lo reparte sola y el caso base no
        # cambia. Con B16 = 25 % (la sensibilidad `delivery_sup`) el IVA
        # repercutido sube con él, que es lo que pide la ley.
        alc, deliv = P('pct_bebida_alc'), P('pct_delivery')
        iva_bebida = ('(' + alc + '*((1-' + deliv + ')*' + P('iva_bebida')
                      + '+' + deliv + '*' + P('iva_general') + ')+(1-' + alc
                      + ')*' + P('iva_reducido') + ')')
        self.iva_bebida = iva_bebida
        iva_bebida_compra = ('(' + alc + '*' + P('iva_general') + '+(1-' + alc
                             + ')*' + P('iva_reducido') + ')')
        self.iva_bebida_compra = iva_bebida_compra

        def pct(clave):
            return (lambda R, k=clave: '=' + R.c(k) + '/' + R.c(
                'ingresos', absoluta=True))

        rej.add(rot='INGRESOS', bold=True)
        rej.add('cub', rot='Cubiertos/día (media)', fmt=motor.FMT_ENT,
                sin_driver=True,
                formulas={'B': '=' + P('cubiertos_dia'),
                          'C': (lambda R: '=' + R.c('cub') + '*(1+' + crec2
                                + ')'),
                          'D': (lambda R: '=' + R.c('cub', 'C') + '*(1+'
                                + crec3 + ')'),
                          'F': '="El crecimiento de los años 2 y 3 está en la '
                               'hoja de Supuestos"'})
        rej.add('ticket', rot='Ticket medio sin IVA', fmt=motor.FMT_EUR,
                sin_driver=True,
                formulas={'B': '=' + P('ticket_medio'),
                          'C': (lambda R: '=' + R.c('ticket')),
                          'D': (lambda R: '=' + R.c('ticket', 'C')),
                          'F': '="Constante en euros del año 1: una subida de '
                               'precios es una decisión aparte"'})
        rej.add('dias', rot='Días de apertura al año', fmt=motor.FMT_ENT,
                sin_driver=True,
                formulas={'B': '=' + P('dias_apertura'),
                          'C': (lambda R: '=' + R.c('dias')),
                          'D': (lambda R: '=' + R.c('dias', 'C'))})
        rej.add('ingresos', rot='INGRESOS TOTALES (sin IVA)', bold=True,
                sin_driver=True,
                fmt=motor.FMT_EUR0, fmt_E=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=IF(' + R.c('cub', c) + '*'
                            + R.c('ticket', c) + '*' + R.c('dias', c)
                            + '=0,"",' + R.c('cub', c) + '*'
                            + R.c('ticket', c) + '*' + R.c('dias', c) + ')'))
                     for col in ('B', 'C', 'D')]
                    # ⚠️ M14 / R22-PAN-18 (2026-09-05). El 1 iba TECLEADO y
                    # todas sus vecinas de la columna E llevan la guarda de
                    # libro vacío: con el libro en blanco la columna quedaba
                    # entera vacía menos esta celda, que seguía enseñando
                    # «100,0 %». Era el último «100 %» falso del libro.
                    + [('E', (lambda R: '=IFERROR(IF(' + R.c('ingresos')
                              + '="","",1),"")'))]))
        lineas = self.lineas_ingreso()
        # RD-23 / RC-06 / RT-13 — el mix de comida y bebida es el segundo
        # driver del libro y la hoja de Instrucciones lo anuncia como dato de
        # mando de «0. Supuestos», pero el input real vivía aquí, en una hoja
        # protegida y a cuarenta filas de distancia. Con dos líneas (A-α) la
        # dependencia se invierte: el peso se teclea en Supuestos y el P&L lo
        # lee. Con más de dos (A-β) el peso sigue en su línea y Supuestos
        # suma.
        self.mix_en_supuestos = (len(lineas) == 2
                                 and lineas[0][2] == 'comida'
                                 and lineas[1][2] == 'bebida')
        # M19 / REF22-BAR-03 / FT22-09 — qué tipos de IVA declaran de verdad
        # las líneas de venta de ESTE libro. La hoja de Instrucciones
        # explicaba el divisor 1,04 (superreducido) en los cinco hermanos,
        # y sólo la panadería tiene una línea al 4 %.
        self.tipos_especiales = set()
        for i, linea in enumerate(lineas):
            (rot, peso, grupo, nota, _fuente,
             tipo_iva) = (list(linea) + [None] * 6)[:6]
            clave = 'lin_%d' % i
            ultimo = i == len(lineas) - 1
            nota_linea = nota or ('Bebida' if grupo == 'bebida' else 'Comida')
            spec = {
                'rot': rot, 'fmt': motor.FMT_EUR0, 'fmt_E': motor.FMT_PCT,
                'fmt_G': motor.FMT_PCT,
                'formulas': dict(
                    (col, (lambda R, c=col, k=clave: '=' + R.c('ingresos', c)
                           + '*' + R.c(k, 'E', absoluta=True)))
                    for col in ('B', 'C', 'D')),
                'valores': {},
            }
            # ⚠️ A3 / REF-04 (2026-09-05). El IVA REPERCUTIDO se calculaba a
            # golpe de dos porcentajes globales —`pct_comida × reducido +
            # pct_bebida × mezcla_de_bebida`— y no había NINGUNA celda donde
            # decir que una línea tributa a otro tipo. El único hermano cuyo
            # producto principal es PAN COMÚN (tipo superreducido del 4 %,
            # art. 91.Dos.1.1.º de la Ley 37/1992) repercutía el 10 % sobre
            # toda su facturación, inflaba el saldo de caja y publicaba un
            # PVP equivalente que no era el de una barra de pan. Ahora el
            # tipo vive EN LA LÍNEA, y los tres consumidores (cobros con IVA,
            # IVA repercutido del mes y PVP equivalente) lo leen con un
            # SUMPRODUCT de pesos × tipos. Sin declarar nada, el resultado es
            # exactamente el de antes.
            if isinstance(tipo_iva, dict):
                self.tipos_especiales.add(
                    round(float(tipo_iva.get('tipo', 0.04)), 4))
                # línea PARTIDA entre dos tipos: el porcentaje que va al tipo
                # especial es un INPUT verde en la propia fila (columna H)
                cabecera_h = (tipo_iva.get('cabecera')
                              or 'Parte de la línea al tipo especial (%)')
                motor.val(ws, 'H' + str(cab), cabecera_h, bold=True,
                          wrap=True)
                spec['valores']['H'] = tipo_iva.get('pct', 0)
                spec['fmt_H'] = motor.FMT_PCT
                # M10 — la primera línea partida es la que fija el peso del
                # producto al tipo especial; `IVA_COMPRAS` puede citarla.
                if getattr(self, 'ref_peso_especial', None) is None:
                    # la fila todavía no está en la rejilla (se añade abajo):
                    # su número es el siguiente al de las ya declaradas
                    self.ref_peso_especial = '$H$' + str(
                        rej.fila0 + len(rej.filas))
                _bajo = repr(float(tipo_iva.get('tipo', 0.04)))
                spec['formulas']['G'] = (
                    lambda R, k=clave, b=_bajo: '=' + R.c(k, 'H') + '*' + b
                    + '+(1-' + R.c(k, 'H') + ')*' + P('iva_reducido'))
                if tipo_iva.get('nota'):
                    nota_linea = nota_linea + ' · ' + tipo_iva['nota']
            elif isinstance(tipo_iva, str) and tipo_iva.startswith('='):
                spec['formulas']['G'] = tipo_iva
            elif tipo_iva is not None:
                self.tipos_especiales.add(round(float(tipo_iva), 4))
                spec['valores']['G'] = float(tipo_iva)
                nota_linea = (nota_linea + ' · El tipo de IVA que REPERCUTE '
                              'esta línea es un input (celda verde de la '
                              'derecha): cámbialo si tu carta tributa a otro '
                              'tipo.')
            else:
                spec['formulas']['G'] = ('=' + (iva_bebida
                                                if grupo == 'bebida'
                                                else P('iva_reducido')))
            spec['valores']['F'] = nota_linea
            if self.mix_en_supuestos:
                spec['formulas']['E'] = (
                    '=' + P('pct_comida' if grupo == 'comida'
                            else 'pct_bebida'))
            elif ultimo and len(lineas) > 1:
                otros = ['lin_%d' % j for j in range(len(lineas) - 1)]
                spec['formulas']['E'] = (lambda R, ks=tuple(otros):
                                         '=1-' + '-'.join(R.c(k, 'E')
                                                          for k in ks))
            else:
                spec['valores']['E'] = peso
                spec['verdes'] = ('E',)
            _verdes = list(spec.get('verdes') or ())
            for _c in ('G', 'H'):
                if _c in spec['valores'] and _c not in _verdes:
                    _verdes.append(_c)
            if _verdes:
                spec['verdes'] = tuple(_verdes)
            rej.add(clave, **spec)
        # A3 — pesos × tipos de las líneas de venta: el IVA repercutido MEDIO
        # del libro, en un solo sitio y calculado desde la propia tabla.
        self.iva_ventas = (
            'SUMPRODUCT(' + rej.r('lin_0', 'E') + ':'
            + rej.c('lin_%d' % (len(lineas) - 1), 'E', absoluta=True) + ','
            + rej.r('lin_0', 'G') + ':'
            + rej.c('lin_%d' % (len(lineas) - 1), 'G', absoluta=True) + ')')
        rej.add(rot='COSTES VARIABLES', bold=True)
        claves_var = []

        def variable(clave, rot, factor, nota=None, iva=None,
                     iva_valor=None):
            """Coste que sube y baja con las ventas.

            `iva` es el tipo que se soporta al comprarlo, y lo lee la hoja de
            Tesorería (RD-13): aplicar un 21 % plano a todo se deducía un IVA
            inexistente en las partidas exentas.
            """
            claves_var.append(clave)
            spec = {'rot': rot, 'fmt': motor.FMT_EUR0,
                    'fmt_E': motor.FMT_PCT, 'fmt_G': motor.FMT_PCT,
                    'formulas': dict(
                        [(col, (lambda R, c=col, f=factor:
                                '=' + R.c('ingresos', c) + '*' + f))
                         for col in ('B', 'C', 'D')]
                        + [('E', pct(clave))])}
            if iva is not None:
                spec['formulas']['G'] = ('=' + iva if not str(iva).startswith(
                    '=') else str(iva))
            if iva_valor is not None:
                # M10 — tipo de compra tecleable: celda VERDE con su valor
                spec.setdefault('valores', {})['G'] = iva_valor
                spec['verdes'] = tuple(set(spec.get('verdes') or ()) | {'G'})
            if nota:
                destino = ('formulas' if str(nota).startswith('=')
                           else 'valores')
                spec.setdefault(destino, {})['F'] = nota
            rej.add(clave, **spec)

        # ⚠️ M10 / R22-PAN-05 (b) (2026-09-05). El IVA que se SOPORTA al
        # comprar la comida iba al 10 % en los cinco hermanos. En una
        # panadería la compra dominante son HARINAS PANIFICABLES, que van al
        # 4 % igual que el pan común (art. 91.Dos.1.1.º de la Ley 37/1992):
        # el libro ya separaba el 4 % en la VENTA y no lo separaba en la
        # COMPRA, así que inflaba el IVA soportado y, como todo queda «a
        # compensar», arrastraba el saldo de caja y el payback. El molde
        # puede sobreescribir el tipo con `IVA_COMPRAS`:
        #   * un NÚMERO → celda verde en la columna del tipo, con su nota;
        #   * una FÓRMULA con `{peso_especial}` (el input verde que ya
        #     declara la línea de venta partida) y `{iva_reducido}`.
        # La parte (a) del hallazgo —tipo de IVA por FILA en la hoja de
        # Inversión— queda DIFERIDA a propósito.
        _ivac = self.dato('IVA_COMPRAS', {}) or {}
        _iva_comida = _ivac.get('comida')
        _nota_comida = texto_pct(P('coste_comida'), 'Al ',
                                 ' de las ventas de COMIDA, no del total: la '
                                 'bebida tiene su propia línea')
        _verde_comida = None
        if isinstance(_iva_comida, str) and _iva_comida.strip().startswith(
                '='):
            _peso = getattr(self, 'ref_peso_especial', None)
            _iva_comida = (_iva_comida
                           .replace('{peso_especial}', _peso or '0')
                           .replace('{iva_reducido}', P('iva_reducido'))
                           .replace('{iva_general}', P('iva_general')))
        elif _iva_comida is not None:
            _verde_comida = float(_iva_comida)
            _iva_comida = None
        if _ivac.get('nota'):
            _nota_comida = (_nota_comida[:-1] + ' · '
                            + str(_ivac['nota']).replace('"', '') + '"')
        variable('cv_comida', 'Coste de mercancía — comida',
                 P('pct_comida') + '*' + P('coste_comida'),
                 nota=_nota_comida,
                 iva=(_iva_comida if _iva_comida is not None
                      else (None if _verde_comida is not None
                            else P('iva_reducido'))),
                 iva_valor=_verde_comida)
        # RD-17 — el tipo que se SOPORTA al comprar la bebida no cambia con
        # la decisión del 10 % en sala: mezcla con `iva_general` (el alcohol
        # se compra al 21 %), no con `iva_bebida`, que es el repercutido.
        variable('cv_bebida', 'Coste de mercancía — bebida',
                 P('pct_bebida') + '*' + P('coste_bebida'),
                 nota=texto_pct(P('coste_bebida'), 'Al ',
                                ' de las ventas de BEBIDA'),
                 iva=iva_bebida_compra)
        variable('cv_cons', 'Consumibles y envases', P('pct_consumibles'),
                 iva=P('iva_soportado'))
        variable('cv_deliv', 'Comisiones de delivery',
                 P('pct_delivery') + '*' + P('comision_delivery'),
                 nota='="Sobre las ventas DEL CANAL. Con 0 % de delivery en '
                      'Supuestos, esta línea vale cero"',
                 iva=P('iva_soportado'))
        variable('cv_tpv', 'Comisiones de los medios de pago',
                 P('comision_tpv'),
                 nota='="Las comisiones de los medios de pago están EXENTAS '
                      'de IVA (art. 20.Uno.18 LIVA): no se deduce nada por '
                      'ellas"',
                 iva='0')
        # RT-04 / RT-05 — «Varios e imprevistos» es un 2 % de las VENTAS y
        # vivía dentro de COSTES FIJOS. Consecuencia medida: los escenarios
        # pesimista y optimista se movían al cambiar el caso realista (los
        # tres importan el mismo bloque de fijos) y el punto de equilibrio
        # subía con el volumen, que es algo que un break-even lineal no puede
        # hacer. Es un coste variable y va donde le toca.
        variable('cv_varios', 'Varios e imprevistos', P('pct_varios'),
                 nota=texto_pct(P('pct_varios'), 'Al ',
                                ' de las ventas. Es un coste VARIABLE: sube '
                                'y baja con la facturación, así que no puede '
                                'vivir en los costes fijos'),
                 iva=P('iva_soportado'))
        rej.add('tcv', rot='TOTAL COSTES VARIABLES', bold=True,
                fmt=motor.FMT_EUR0, fmt_E=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=SUM(' + R.c(claves_var[0], c)
                            + ':' + R.c(claves_var[-1], c) + ')'))
                     for col in ('B', 'C', 'D')] + [('E', pct('tcv'))]))
        self.claves_var = claves_var
        rej.add('mb', rot='MARGEN BRUTO', bold=True, fmt=motor.FMT_EUR0,
                fmt_E=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=' + R.c('ingresos', c) + '-'
                            + R.c('tcv', c)))
                     for col in ('B', 'C', 'D')] + [('E', pct('mb'))]))
        rej.add(rot='COSTES FIJOS', bold=True)
        claves_fijos = []

        def fijo(clave, rot, formula_b, nota=None, verde=False, valor=None,
                 iva=None):
            # A6 / M-05 — «Alquiler del local» en un negocio sin local. Las
            # filas canónicas las escribe este grupo, así que su rótulo no
            # pasaba por `_fijos` y el módulo de contenido no podía tocarlo.
            _acc = self._reglas('FIJOS').get(motor.norm(rot))
            if _acc and len(_acc) > 2 and _acc[2]:
                rot = _acc[2]
            if _acc and _acc[0] is None and _acc[1]:
                nota = _acc[1]
            claves_fijos.append(clave)
            spec = {'rot': rot, 'fmt': motor.FMT_EUR0,
                    'fmt_E': motor.FMT_PCT, 'fmt_G': motor.FMT_PCT,
                    'formulas': {'E': pct(clave)}}
            if valor is not None:
                spec['valores'] = {'B': valor}
                spec['verdes'] = ('B',)     # la nota no se pinta de verde
            else:
                spec['formulas']['B'] = formula_b
            factor = '*(1+' + ipc + ')'
            spec['formulas']['C'] = (lambda R, k=clave, f=factor:
                                     '=' + R.c(k) + f)
            spec['formulas']['D'] = (lambda R, k=clave, f=factor:
                                     '=' + R.c(k, 'C') + f)
            if iva is not None:
                spec['formulas']['G'] = '=' + iva
            if nota:
                destino = ('formulas' if str(nota).startswith('=')
                           else 'valores')
                spec.setdefault(destino, {})['F'] = nota
            rej.add(clave, **spec)

        fijo('cf_alquiler', 'Alquiler del local',
             '=' + P('alquiler_mes') + '*12',
             nota='="El importe mensual está en la hoja de Supuestos; aquí '
                  'se multiplica por doce. La renta de los meses de obra, '
                  'que es anterior a la apertura, está en la hoja de '
                  'Inversión y NO se cuenta dos veces"',
             iva=P('iva_soportado'))
        fijo('cf_personal', 'Personal (nóminas + Seguridad Social)',
             (lambda R: '=' + self.rej['personal'].r('total', 'G')),
             nota='="Sale de la hoja de Personal: es el MISMO número, no una '
                  'estimación aparte. Los años 2 y 3 suman las '
                  'INCORPORACIONES declaradas en esa hoja, no un porcentaje '
                  'de volumen"',
             iva='0')
        # RD-10 — con el personal creciendo al ritmo del volumen, el ratio de
        # coste laboral salía IDÉNTICO en los tres años por construcción y el
        # semáforo de los años 2 y 3 no podía fallar nunca: un banco leía
        # tres años de ratios donde sólo había uno.
        rej.filas[rej.fila('cf_personal') - rej.fila0]['formulas']['C'] = (
            lambda R: '=' + self.rej['personal'].r('coste_2'))
        rej.filas[rej.fila('cf_personal') - rej.fila0]['formulas']['D'] = (
            lambda R: '=' + self.rej['personal'].r('coste_3'))
        # FIS-07: el agua apta para el consumo humano va al 10 % como
        # entrega de bienes; la partida mezcla luz, agua y gas al general.
        fijo('cf_suministros', 'Suministros (luz, agua, gas)',
             '=' + P('suministros_mes') + '*12', iva=P('iva_soportado'),
             nota='Luz y gas van al 21 %; la factura del agua, al 10 %. El '
                  'libro aplica a toda esta partida el tipo de la columna de '
                  'al lado (21 %): si el agua pesa mucho en tu local, '
                  'sepárala en su línea')
        # RD-13 — los seguros están EXENTOS (art. 20.Uno.16 LIVA): el modelo
        # se deducía un IVA que no existe y con él bajaba la liquidación.
        fijo('cf_seguros', 'Seguros (RC + multirriesgo)',
             '=' + P('seguros_ano'),
             nota='="Las operaciones de seguro están EXENTAS de IVA (art. '
                  '20.Uno.16 LIVA): la prima no lleva IVA deducible"',
             iva='0')
        for i, (rot, importe, nota) in enumerate(fijos):
            fijo('cf_p%d' % i, rot, None, valor=importe,
                 nota=nota if isinstance(nota, str) else None,
                 iva=P('iva_soportado'))
        fijo('cf_amort', 'Amortización del inmovilizado',
             (lambda R: '=' + self.rej['inversion'].r('amort')),
             nota='="Sale de las bases de amortización de la hoja de '
                  'Inversión: sólo inmovilizado real. NO sube con el IPC: es '
                  'una cuota sobre el coste HISTÓRICO del inmovilizado"',
             iva='0')
        # RT-21 — la amortización se actualizaba con la inflación de costes.
        # Es una cuota sobre el coste histórico: no sube con el IPC. El
        # defecto estaba latente porque `ipc` vale 0, pero la celda es verde
        # y su propia nota invita a subirla; con un 2 % la amortización del
        # año 3 salía un 4,04 % por encima de la base y arrastraba al
        # resultado, al DSCR y al flujo de caja libre.
        for col in ('C', 'D'):
            rej.filas[rej.fila('cf_amort') - rej.fila0]['formulas'][col] = (
                lambda R, c=col: '=' + R.c('cf_amort'))
        fijo('cf_int', 'Gastos financieros (intereses del préstamo)',
             (lambda R: '=' + self.rej['financiacion'].r('int_1')),
             nota='="Sólo los INTERESES son gasto. La devolución del '
                  'principal va en la hoja de Tesorería"',
             iva='0')
        rej.filas[rej.fila('cf_int') - rej.fila0]['formulas']['C'] = (
            lambda R: '=' + self.rej['financiacion'].r('int_2'))
        rej.filas[rej.fila('cf_int') - rej.fila0]['formulas']['D'] = (
            lambda R: '=' + self.rej['financiacion'].r('int_3'))
        self.claves_fijos = claves_fijos
        rej.add('tcf', rot='TOTAL COSTES FIJOS', bold=True,
                fmt=motor.FMT_EUR0, fmt_E=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=SUM(' + R.c(claves_fijos[0], c)
                            + ':' + R.c(claves_fijos[-1], c) + ')'))
                     for col in ('B', 'C', 'D')] + [('E', pct('tcf'))]))
        # RD-13 — IVA soportado del año por bloque, leyendo la columna G:
        # la tesorería ya no aplica un 21 % plano a partidas exentas.
        rej.add('iva_var', rot='IVA soportado de los costes variables',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(col, (lambda R, c=col:
                            '=SUMPRODUCT(' + R.c(claves_var[0], c) + ':'
                            + R.c(claves_var[-1], c) + ','
                            + R.c(claves_var[0], 'G') + ':'
                            + R.c(claves_var[-1], 'G') + ')'))
                     for col in ('B', 'C', 'D')]
                    + [('F', '="Cada línea a SU tipo, leído de la columna de '
                        'al lado. Es lo que la hoja de Tesorería paga de más '
                        'a los proveedores y luego se deduce"')]))
        rej.add('iva_fij', rot='IVA soportado de los costes fijos',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col:
                           '=SUMPRODUCT(' + R.c(claves_fijos[0], c) + ':'
                           + R.c(claves_fijos[-1], c) + ','
                           + R.c(claves_fijos[0], 'G') + ':'
                           + R.c(claves_fijos[-1], 'G') + ')'))
                    for col in ('B', 'C', 'D')))
        rej.add('rai', rot='RESULTADO ANTES DE IMPUESTOS', bold=True,
                fmt=motor.FMT_EUR0, fmt_E=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=' + R.c('mb', c) + '-'
                            + R.c('tcf', c)))
                     for col in ('B', 'C', 'D')] + [('E', pct('rai'))]))
        # Impuesto de Sociedades con BIN y tipo de nueva creación (TEC-06)
        rej.add('bin_ini', rot='Bases negativas pendientes al inicio',
                fmt=motor.FMT_EUR0,
                formulas={'B': '=' + P('bin_inicial'),
                          'C': (lambda R: '=' + R.c('bin_fin')),
                          'D': (lambda R: '=' + R.c('bin_fin', 'C')),
                          'F': '="Art. 26 LIS: las pérdidas de un ejercicio '
                               'se compensan con los beneficios de los '
                               'siguientes"'})
        rej.add('base', rot='Base imponible (después de compensar)',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=MAX(0,' + R.c('rai', c) + '-'
                           + R.c('bin_ini', c) + ')'))
                    for col in ('B', 'C', 'D')))
        rej.add('acum', rot='Ejercicios con base positiva (acumulado)',
                fmt=motor.FMT_ENT,
                formulas={'B': (lambda R: '=IF(' + R.c('base') + '>0,1,0)'),
                          'C': (lambda R: '=' + R.c('acum') + '+IF('
                                + R.c('base', 'C') + '>0,1,0)'),
                          'D': (lambda R: '=' + R.c('acum', 'C') + '+IF('
                                + R.c('base', 'D') + '>0,1,0)')})
        rej.add('tipo', rot='Tipo de Impuesto de Sociedades aplicado',
                fmt=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=IF(' + R.c('base', c)
                            + '<=0,"",IF(' + R.c('acum', c) + '<=2,'
                            + P('is_nueva') + ',' + P('is_general') + '))'))
                     for col in ('B', 'C', 'D')]
                    + [('F', '="Art. 29.1 LIS: el 15 % de entidad de nueva '
                        'creación se aplica al PRIMER ejercicio con base '
                        'positiva y al SIGUIENTE, no al primer año de vida. '
                        'El libro lo aproxima contando los dos primeros '
                        'ejercicios con base positiva"')]))
        rej.add('is', rot='Impuesto de Sociedades', fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=IF(' + R.c('tipo', c)
                           + '="",0,' + R.c('base', c) + '*' + R.c('tipo', c)
                           + ')'))
                    for col in ('B', 'C', 'D')))
        rej.add('bin_fin', rot='Bases negativas pendientes al cierre',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=MAX(0,' + R.c('bin_ini', c)
                           + '-' + R.c('rai', c) + ')'))
                    for col in ('B', 'C', 'D')))
        rej.add('neto', rot='RESULTADO NETO', bold=True, fmt=motor.FMT_EUR0,
                fmt_E=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=' + R.c('rai', c) + '-'
                            + R.c('is', c)))
                     for col in ('B', 'C', 'D')] + [('E', pct('neto'))]))
        # ---- ratios que auditan (TEC-12, §2.9) --------------------------
        rej.add(rot='RATIOS CLAVE', bold=True)
        rej.add('cab_ratios', rot='Ratio', bold=True,
                valores={'B': 'Año 1', 'C': 'Año 2', 'D': 'Año 3',
                         'E': 'Umbral', 'F': 'Comentario'})
        umbrales = dict((u[0], u) for u in (self.dato('UMBRALES') or ()))

        def ratio(clave, rot, numerador, umbral, comentario, verde_si_menor,
                  fmt_um=motor.FMT_PCT):
            u = umbrales.get(clave)
            valor = u[2] if u else umbral
            texto = u[3] if u else comentario
            # A6 / M-05 — hasta ahora `ratio()` sólo tomaba de la tupla el
            # VALOR y el COMENTARIO: el rótulo que el módulo de contenido
            # declaraba era letra muerta y el libro publicaba «Alquiler /
            # Ventas» en un producto sin local.
            rot = (u[1] if u and u[1] else rot)
            rej.add(clave, rot=rot, fmt=motor.FMT_PCT, fmt_E=fmt_um,
                    valores={'E': valor, 'F': texto}, verdes=('E',),
                    formulas=dict(
                        (col, (lambda R, c=col, n=numerador:
                               '=' + R.c(n, c) + '/' + R.c('ingresos', c)))
                        for col in ('B', 'C', 'D')))
            ancla = rej.c(clave)
            comparador = ('<=' if verde_si_menor else '>=')
            contrario = ('>' if verde_si_menor else '<')
            motor.semaforo_num(
                rej.ws, ancla + ':' + rej.c(clave, 'D'),
                verde_si=ancla + comparador + rej.c(clave, 'E', absoluta=True),
                rojo_si=ancla + contrario + rej.c(clave, 'E', absoluta=True))

        ratio('r_mb', 'Margen bruto / Ventas', 'mb', 0.65,
              'Objetivo del formato; por debajo, revisa precios o compras',
              False)
        rej.add('cogs', rot='Coste de mercancía (comida + bebida)',
                fmt=motor.FMT_EUR0, fmt_E=motor.FMT_PCT,
                formulas=dict(
                    (col, (lambda R, c=col: '=' + R.c('cv_comida', c) + '+'
                           + R.c('cv_bebida', c)))
                    for col in ('B', 'C', 'D')))
        ratio('r_cogs', 'Coste de mercancía / Ventas', 'cogs', 0.32,
              'El food cost REAL del libro, calculado, no tecleado', True)
        ratio('r_personal', 'Coste de personal / Ventas', 'cf_personal', 0.35,
              'Techo que fijan las Instrucciones de este mismo libro', True)
        ratio('r_alquiler', 'Alquiler / Ventas', 'cf_alquiler', 0.10,
              'Por encima del 10 % el local se come el margen', True)
        ratio('r_neto', 'Resultado neto / Ventas', 'neto', 0.05,
              'Suelo de rentabilidad del sector', False)
        # RD-20 — la fila «Punto de equilibrio alcanzado» comparaba los años
        # 2 y 3 contra el umbral del AÑO 1, que se calcula con los costes
        # fijos del año 1. Como el personal crece, el umbral real de los años
        # 2 y 3 es mayor y la respuesta «Sí» estaba dada contra una
        # referencia caducada: ahora cada año se mide contra el suyo.
        rej.add('r_be', rot='Punto de equilibrio alcanzado',
                formulas=dict(
                    [(col, (lambda R, c=col: '=IF(' + R.c('ingresos', c)
                            + '="","",IF(' + R.c('ingresos', c) + '>='
                            + self.rej['equilibrio'].r('ingresos_be', c)
                            + ',"Sí","No"))')) for col in ('B', 'C', 'D')]
                    + [('F', '="Compara las ventas de CADA año con el umbral '
                        'de ese mismo año, que calcula la hoja de Punto de '
                        'Equilibrio con los costes fijos del año"')]))
        motor.semaforo_texto(rej.ws, rej.c('r_be') + ':' + rej.c('r_be', 'D'),
                             (('Sí', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
                              ('No', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
        self.pendientes.append(rej)
        fila = rej.ultima + 2
        motor.val(ws, 'A' + str(fila),
                  'Todas las cifras van SIN IVA. Las columnas de los años 2 '
                  'y 3 están en euros del año 1 salvo que subas la '
                  'actualización de costes en la hoja de Supuestos. La '
                  'columna «Tipo de IVA de la línea» dice, en las líneas de '
                  'VENTA, a qué tipo se repercute cada una, y en las de '
                  'COSTE a qué tipo se compra la partida. La lee la hoja de '
                  'Tesorería: los seguros y las comisiones de los medios de '
                  'pago están exentos, y las nóminas, la amortización y los '
                  'intereses no llevan IVA.',
                  wrap=True)
        for i, texto in enumerate(pie):
            motor.val(ws, 'A' + str(fila + 1 + i), texto)
        motor.anchos(ws, {'A': 42, 'B': 15, 'C': 15, 'D': 15, 'E': 13,
                          'F': 60, 'G': 14, 'H': 16})
        motor.print_setup(ws, header_row=cab)
        return rej


    def _fijos(self, ws, cab):
        """Costes fijos que se CONSERVAN como input (§1.3)."""
        reglas = self._reglas('FIJOS')
        fuera = []
        # SOLO el bloque de costes fijos: debajo están el EBITDA, el impuesto
        # y el resultado neto, que también son números en la columna B.
        r0, r1 = _seccion(ws, cab, r'^costes fijos',
                          r'^total costes fijos')
        if r0 is None:
            self.anota('P&L: no se encuentra el bloque «COSTES FIJOS» en '
                       + ws.title + ': sólo se escriben las partidas del '
                       'módulo de contenido (no se adivina)')
            r0 = r1 = cab
        for _r, rot, importe, nota, _bloque in _partidas(
                ws, cab, canon=RX_CANON_PYG, desde=r0, hasta=r1):
            rot, cambiado = _limpiar_rotulo(rot)
            if cambiado:
                self.anota('P&L: rótulo con parámetro escrito a mano → «'
                           + rot + '» (§7-bis.11)')
            accion = reglas.get(motor.norm(rot))
            if accion and accion[0] == 'suprimir':
                self.anota('P&L: fuera «' + rot + '» — '
                           + (accion[1] or 'lo pide el módulo de contenido'))
                continue
            if accion and motor._es_numero(accion[0]):
                self.anota('P&L: «' + rot + '» ' + str(importe) + ' → '
                           + str(accion[0]) + ' — ' + (accion[1] or ''))
                importe, nota = float(accion[0]), accion[1] or nota
            elif accion and accion[0] is None and accion[1]:
                nota = accion[1]
            if accion and len(accion) > 2 and accion[2] \
                    and motor.norm(accion[2]) != motor.norm(rot):
                self.anota('P&L: rótulo «' + rot + '» → «' + accion[2]
                           + '» (A6)')
                rot = accion[2]
            fuera.append((rot, importe, nota))
        # los extras se añaden UNA vez: en la 2.ª pasada ya están dentro del
        # bloque de costes fijos y `_partidas` los devuelve como preservados
        ya = set(motor.norm(r) for r, _i, _n in fuera)
        for extra in (self.dato('FIJOS_EXTRA', []) or []):
            rot, importe, nota, _fuente = (list(extra) + [None] * 4)[:4]
            if motor.norm(rot) in ya:
                continue
            ya.add(motor.norm(rot))
            fuera.append((rot, importe, nota))
        return fuera

    # -- §2.4 -------------------------------------------------------------
    def equilibrio(self):
        """`Punto de Equilibrio`: contable y de caja, por año (§2.4)."""
        ws = self.ws_equilibrio
        cab = _cabecera(ws)
        pie = _pie(ws, cab)
        _limpiar_area(ws, cab, ws.max_row + 20, 6)
        for i, texto in enumerate(('Variable', 'Año 1', 'Año 2', 'Año 3',
                                   'Notas')):
            motor.val(ws, get_column_letter(i + 1) + str(cab), texto,
                      bold=True)
        pyg = self.rej['pyg']
        rej = Rejilla(ws, cab + 1,
                      driver=lambda R, c, k: pyg.r('ingresos', c))
        self.rej['equilibrio'] = rej
        fin = self.rej['financiacion']
        cols = ('B', 'C', 'D')
        rej.add(rot='DATOS BASE POR AÑO', bold=True)
        rej.add('cf_ano', rot='Costes fijos anuales', fmt=motor.FMT_EUR,
                formulas=dict(
                    [(c, '=' + pyg.r('tcf', c)) for c in cols]
                    + [('E', '="Los MISMOS costes fijos del P&L. Cada año '
                        'tiene los suyos: por eso el umbral no puede ser uno '
                        'solo para los tres"')]))
        rej.add('cf_mes', rot='Costes fijos mensuales', fmt=motor.FMT_EUR,
                formulas=dict(
                    (c, (lambda R, cc=c: '=' + R.c('cf_ano', cc) + '/12'))
                    for c in cols))
        rej.add('ticket', rot='Ticket medio sin IVA', fmt=motor.FMT_EUR,
                formulas=dict((c, '=' + pyg.r('ticket', c)) for c in cols))
        rej.add('cvu', rot='Coste variable por cubierto', fmt=motor.FMT_EUR,
                formulas=dict(
                    [(c, '=' + pyg.r('tcv', c) + '/(' + pyg.r('cub', c) + '*'
                      + pyg.r('dias', c) + ')') for c in cols]
                    + [('E', '="Sale del total de costes variables del P&L, '
                        'no de una estimación aparte. Incluye los varios e '
                        'imprevistos, que son un porcentaje de las ventas"')]))
        rej.add('mc', rot='Margen de contribución por cubierto',
                fmt=motor.FMT_EUR,
                formulas=dict(
                    [(c, (lambda R, cc=c: '=' + R.c('ticket', cc) + '-'
                          + R.c('cvu', cc))) for c in cols]
                    + [('E', '="Ticket menos coste variable unitario"')]))
        rej.add('dias', rot='Días de apertura al año', fmt=motor.FMT_ENT,
                formulas=dict(
                    [(c, '=' + pyg.r('dias', c)) for c in cols]
                    + [('E', '="El mismo calendario que el P&L y que los '
                        'escenarios"')]))
        # RD-05 — el break-even excluía la CUOTA, contra la decisión
        # §7-bis.2 («break-even ÚNICO que incluye amortización y cuota»). Los
        # costes fijos llevan amortización e intereses pero no el principal,
        # así que el umbral publicado estaba por debajo del real. Se entregan
        # los DOS umbrales rotulados, como manda §3.5.
        rej.add('principal', rot='Devolución de principal del año (cuota)',
                fmt=motor.FMT_EUR,
                formulas=dict(
                    [(c, '=' + fin.r('cap_%d' % (i + 1)))
                     for i, c in enumerate(cols)]
                    + [('E', '="El principal NO es gasto del P&L, pero sale '
                        'de la caja. Sin él, el umbral de equilibrio se '
                        'publica por debajo del real"')]))
        rej.add(rot='PUNTO DE EQUILIBRIO CONTABLE (incluye amortización)',
                bold=True)
        rej.add('cub_ano', rot='Cubiertos necesarios al año',
                fmt=motor.FMT_ENT,
                formulas=dict(
                    (c, (lambda R, cc=c: '=IF(' + R.c('mc', cc)
                         + '<=0,"",' + R.c('cf_ano', cc) + '/'
                         + R.c('mc', cc) + ')')) for c in cols))
        rej.add('cub_dia', rot='Cubiertos necesarios al día',
                fmt=motor.FMT_DEC,
                formulas=dict(
                    (c, (lambda R, cc=c: '=' + R.c('cub_ano', cc) + '/'
                         + R.c('dias', cc))) for c in cols))
        rej.add('ingresos_be', rot='Ingresos necesarios al año',
                fmt=motor.FMT_EUR0, bold=True,
                formulas=dict(
                    (c, (lambda R, cc=c: '=' + R.c('cub_ano', cc) + '*'
                         + R.c('ticket', cc))) for c in cols))
        rej.add(rot='PUNTO DE EQUILIBRIO DE CAJA (incluye la cuota del '
                'préstamo)', bold=True)
        rej.add('cf_caja', rot='Costes fijos de caja + principal del año',
                fmt=motor.FMT_EUR,
                formulas=dict(
                    [(c, (lambda R, cc=c: '=' + R.c('cf_ano', cc) + '-'
                          + pyg.r('cf_amort', cc) + '+'
                          + R.c('principal', cc))) for c in cols]
                    + [('E', '="Los fijos del P&L sin la amortización (que no '
                        'se paga) y CON la devolución de principal (que sí). '
                        'Es el umbral que decide si la caja aguanta"')]))
        rej.add('cub_ano_caja', rot='Cubiertos necesarios al año (caja)',
                fmt=motor.FMT_ENT,
                formulas=dict(
                    (c, (lambda R, cc=c: '=IF(' + R.c('mc', cc)
                         + '<=0,"",' + R.c('cf_caja', cc) + '/'
                         + R.c('mc', cc) + ')')) for c in cols))
        rej.add('cub_dia_caja', rot='Cubiertos necesarios al día (caja)',
                fmt=motor.FMT_DEC, bold=True,
                formulas=dict(
                    (c, (lambda R, cc=c: '=' + R.c('cub_ano_caja', cc) + '/'
                         + R.c('dias', cc))) for c in cols))
        rej.add('ingresos_be_caja', rot='Ingresos necesarios al año (caja)',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    (c, (lambda R, cc=c: '=' + R.c('cub_ano_caja', cc) + '*'
                         + R.c('ticket', cc))) for c in cols))
        rej.add(rot='CONTRASTE CON EL PLAN', bold=True)
        rej.add('cub_plan', rot='Cubiertos/día previstos en el plan',
                fmt=motor.FMT_ENT,
                formulas=dict((c, '=' + pyg.r('cub', c)) for c in cols))
        holgura_min = self.p.ref('holgura_min')
        rej.add('holgura', rot='Holgura sobre el equilibrio de CAJA (%)',
                fmt=motor.FMT_PCT,
                formulas=dict(
                    [(c, (lambda R, cc=c: '=' + R.c('cub_plan', cc) + '/'
                          + R.c('cub_dia_caja', cc) + '-1')) for c in cols]
                    + [('E', '="Cuánto puedes caer antes de no poder pagar '
                        'los costes fijos y la cuota. El mínimo exigible '
                        'está en la hoja de Supuestos"')]))
        # RD-22 — el umbral vivía DENTRO de la fórmula del formato
        # condicional, contra §7-bis.11: el comprador no podía ajustarlo sin
        # editar reglas de formato. Ahora sale de una celda de Supuestos.
        motor.semaforo_num(ws, rej.c('holgura') + ':' + rej.c('holgura', 'D'),
                           verde_si=rej.c('holgura') + '>=' + holgura_min,
                           ambar_si=rej.c('holgura') + '>=0',
                           rojo_si=rej.c('holgura') + '<0')
        # RT-28 — la guarda de «margen de contribución negativo» era código
        # muerto: con todos los costes variables definidos como porcentaje de
        # las ventas, el margen unitario es una fracción fija del ticket y no
        # puede ser negativo mientras el ticket sea positivo. La guarda que
        # SÍ puede saltar es la de capacidad: un punto de equilibrio que
        # exige más rotaciones de las que da el local es un plan inviable, y
        # eso el modelo lo permite sin decir nada.
        aforo = self.p.ref('aforo')
        # ⚠️ A5 / REF-01 / M-07 / MOT-02 (2026-09-05). El techo estaba
        # CABLEADO a 3,0 servicios por plaza y día —la magnitud de un
        # comedor— y el contenido no tenía ninguna clave para tocarlo: la
        # panadería entregaba TRES CELDAS ROJAS en el caso base (10,3 / 11,6 /
        # 12,0 rotaciones de un mostrador contra un techo de comedor) y el
        # food truck sólo se salvaba calibrando el aforo, que es otra cosa.
        # Ahora el molde puede fijar su techo o APAGAR el bloque entero
        # cuando su driver es la transacción de mostrador y no el cubierto
        # sentado, que es lo que el propio libro ya declara en Supuestos!C51.
        rota = self.dato('ROTACION', {}) or {}
        if rota.get('activa', True):
            rej.add('rot_max',
                    rot='Rotaciones al día que da el local como máximo',
                    fmt=motor.FMT_DEC, verdes=('B',),
                    valores={'B': _num(rota.get('max'), 3.0),
                             'E': rota.get('nota_max')
                             or 'Servicios completos por plaza y día que '
                                'puedes llegar a hacer con tu horario. Es el '
                                'techo físico del local, no un objetivo'})
            rej.add('rot_be', rot='Rotaciones al día que exige el equilibrio '
                    'de caja', fmt=motor.FMT_DEC,
                    formulas=dict(
                        [(c, (lambda R, cc=c: '=' + R.c('cub_dia_caja', cc)
                              + '/' + aforo)) for c in cols]
                        + [('E', '="Cubiertos de equilibrio entre las plazas '
                            'que declara la hoja de Supuestos. Por encima del '
                            'techo de arriba, el punto de equilibrio está '
                            'fuera del alcance del local"')]))
            motor.semaforo_num(ws, rej.c('rot_be') + ':'
                               + rej.c('rot_be', 'D'),
                               verde_si=rej.c('rot_be') + '<='
                               + rej.c('rot_max', absoluta=True),
                               rojo_si=rej.c('rot_be') + '>'
                               + rej.c('rot_max', absoluta=True))
        else:
            self.anota('Punto de equilibrio: bloque de ROTACIONES suprimido '
                       '(ROTACION[\'activa\'] = False): en este molde el '
                       'driver es la transacción de mostrador, no el cubierto '
                       'sentado, y comparar contra un techo de comedor pinta '
                       'de rojo un caso base correcto')
        rej.add(rot='INTERPRETACIÓN', bold=True)
        # RC-30 — el párrafo redondeaba a entero mientras la celda de la que
        # habla muestra un decimal: el cliente leía «67 cubiertos al día» dos
        # líneas debajo de una celda que ponía 67,4. Se redondea HACIA
        # ARRIBA, que es lo defendible: con 67,4 no se cubren los costes.
        # ⚠️ M1 / R22-CAF-03 / R22-PAN-12 (2026-09-05). La frase decía «con
        # la cuota 80, sin la cuota 84»: leída literalmente, quitar un coste
        # SUBÍA el umbral, y es la celda que más se lee de la hoja. La
        # diferencia real no es «con cuota / sin cuota» sino CAJA frente a
        # CONTABLE: el umbral de caja quita la amortización (que no se paga)
        # y suma el principal del año, que en el año 1 vale 0 si el préstamo
        # está en carencia. Ahora los dos umbrales se nombran por lo que son
        # y la carencia se explica sólo cuando la hay (IF sobre la fila del
        # principal, que ya está calculada arriba).
        rej.add('texto', wrap=True, alto=46,
                formulas={'A': (lambda R:
                                '="Con el ticket medio sin IVA de la fila de '
                                'arriba necesitas "&TEXT(ROUNDUP('
                                + R.c('cub_dia') + ',0),"0")&" cubiertos al '
                                'día durante los "&TEXT(' + R.c('dias')
                                + ',"0")&" días que abres para no perder '
                                'dinero (equilibrio contable, con la '
                                'amortización dentro), y "&TEXT(ROUNDUP('
                                + R.c('cub_dia_caja') + ',0),"0")&" para que '
                                'la caja aguante (equilibrio de caja: sin la '
                                'amortización, que no se paga, y con la cuota '
                                'del préstamo del año"&IF(' + R.c('principal')
                                + '=0,"; este año sólo intereses, porque el '
                                'préstamo está en carencia","")&"). En las '
                                'filas «Ingresos necesarios al año» tienes '
                                'esas mismas cifras en euros."')})
        self.pendientes.append(rej)
        fila = self._sensibilidad(ws, rej) + 2
        for i, texto in enumerate(pie):
            motor.val(ws, 'A' + str(fila + i), texto)
        motor.anchos(ws, {'A': 44, 'B': 16, 'C': 16, 'D': 16, 'E': 52,
                          'F': 14})
        motor.print_setup(ws, header_row=cab)
        return rej


    def _sensibilidad(self, ws, rej):
        """Tabla de sensibilidad al ticket y al coste variable (TEC-19)."""
        fila0 = rej.ultima + 2
        motor.val(ws, 'A' + str(fila0), 'SENSIBILIDAD DEL PUNTO DE '
                  'EQUILIBRIO', bold=True)
        motor.val(ws, 'A' + str(fila0 + 1),
                  'Cubiertos/día necesarios para cubrir costes según el '
                  'ticket medio (columnas) y el coste variable por cubierto '
                  '(filas). En verde, los que están por debajo de los '
                  'cubiertos que prevé el plan.', wrap=True)
        ws.row_dimensions[fila0 + 1].height = 30
        cabf = fila0 + 3
        # B12 / R-23 — las dos magnitudes de la tabla van en euros y sin
        # rótulo en el eje de columnas se podía leer girada: A33 rotulaba las
        # FILAS y la cabecera B33:E33 llevaba los TICKETS sin nombre.
        motor.val(ws, 'A' + str(cabf),
                  'Cubiertos/día necesarios: ticket medio sin IVA '
                  '(columnas) × coste variable por cubierto (filas)',
                  bold=True, wrap=True)
        tk = rej.c('ticket', absoluta=True)
        for j, delta in enumerate((-2, 0, 2, 4)):
            col = get_column_letter(2 + j)
            # RT-23 — concatenar «+» y un delta negativo producía `$B$9+-2`:
            # Excel la acepta y pycel la evalúa, pero es una fórmula que el
            # comprador ve al pinchar la celda y que delata el generador.
            fx(ws, col + str(cabf),
               '=' + tk + ('%+g' % delta if delta else ''), motor.FMT_EUR)
            ws[col + str(cabf)].font = Font(bold=True)
        cv = rej.c('cvu', absoluta=True)
        cf = rej.c('cf_ano', absoluta=True)
        dias = rej.c('dias', absoluta=True)
        plan = rej.c('cub_plan', absoluta=True)
        for i, factor in enumerate((0.85, 0.925, 1.0, 1.075, 1.15)):
            fila = cabf + 1 + i
            fx(ws, 'A' + str(fila), '=' + cv + '*' + str(factor),
               motor.FMT_EUR)
            for j in range(4):
                col = get_column_letter(2 + j)
                cabecera = '$' + col + '$' + str(cabf)
                fx(ws, col + str(fila),
                   '=IF(' + cabecera + '-$A' + str(fila) + '<=0,"",' + cf
                   + '/(' + cabecera + '-$A' + str(fila) + ')/' + dias + ')',
                   motor.FMT_DEC)
        rango = 'B' + str(cabf + 1) + ':E' + str(cabf + 5)
        motor.semaforo_num(ws, rango, verde_si='B' + str(cabf + 1) + '<='
                           + plan, rojo_si='B' + str(cabf + 1) + '>' + plan)
        return cabf + 5

    # -- §2.5 -------------------------------------------------------------
    def escenarios(self):
        """`Escenarios`: el MISMO motor que el P&L (§2.5, TEC-02)."""
        ws = self.ws_escenarios
        cab = _cabecera(ws)
        pie = _pie(ws, cab)
        anteriores = self._leer_escenarios(ws, cab)
        _limpiar_area(ws, cab, ws.max_row, 6)
        for i, texto in enumerate(('Métrica', 'Pesimista', 'Realista',
                                   'Optimista', 'Notas')):
            motor.val(ws, get_column_letter(i + 1) + str(cab), texto,
                      bold=True)
        rej = Rejilla(ws, cab + 1,
                      driver=lambda R, c, k: R.c('ingresos', c))
        self.rej['escenarios'] = rej
        pyg = self.rej['pyg']
        P = self.p.ref
        cols = ('B', 'C', 'D')

        def driver(clave, rot, ref_pyg, fmt, idx):
            rej.add(clave, rot=rot, fmt=fmt, verdes=('B', 'D'),
                    sin_driver=True,
                    valores={'B': anteriores['pesimista'][idx],
                             'D': anteriores['optimista'][idx]},
                    formulas={'C': '=' + ref_pyg})

        driver('cub', 'Cubiertos/día', pyg.r('cub'), motor.FMT_ENT, 0)
        driver('ticket', 'Ticket medio sin IVA', pyg.r('ticket'),
               motor.FMT_EUR, 1)
        driver('dias', 'Días de apertura al año', pyg.r('dias'),
               motor.FMT_ENT, 2)
        rej.add('ingresos', rot='INGRESOS ANUALES (sin IVA)', bold=True,
                fmt=motor.FMT_EUR0, sin_driver=True,
                formulas=dict(
                    (col, (lambda R, c=col: '=IF(' + R.c('cub', c) + '*'
                           + R.c('ticket', c) + '*' + R.c('dias', c)
                           + '=0,"",' + R.c('cub', c) + '*'
                           + R.c('ticket', c) + '*' + R.c('dias', c) + ')'))
                    for col in cols))
        # RT-04 — «Varios e imprevistos» entra aquí como coste VARIABLE de
        # cada escenario. Mientras vivía dentro de los costes fijos del P&L,
        # el pesimista y el optimista se movían al cambiar el realista: el
        # pesimista cargaba 8.531 € de imprevistos cuando le tocaban 5.533.
        for clave, rot, factor in (
                ('cv_comida', 'Coste de mercancía — comida',
                 P('pct_comida') + '*' + P('coste_comida')),
                ('cv_bebida', 'Coste de mercancía — bebida',
                 P('pct_bebida') + '*' + P('coste_bebida')),
                ('cv_cons', 'Consumibles y envases', P('pct_consumibles')),
                ('cv_deliv', 'Comisiones de delivery',
                 P('pct_delivery') + '*' + P('comision_delivery')),
                ('cv_tpv', 'Comisiones de los medios de pago',
                 P('comision_tpv')),
                ('cv_varios', 'Varios e imprevistos', P('pct_varios'))):
            rej.add(clave, rot=rot, fmt=motor.FMT_EUR0,
                    formulas=dict(
                        (col, (lambda R, c=col, f=factor:
                               '=' + R.c('ingresos', c) + '*' + f))
                        for col in cols))
        rej.add('tcv', rot='TOTAL COSTES VARIABLES', fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=SUM(' + R.c('cv_comida', c)
                           + ':' + R.c('cv_varios', c) + ')'))
                    for col in cols))
        rej.add('mb', rot='MARGEN BRUTO', fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=' + R.c('ingresos', c) + '-'
                           + R.c('tcv', c))) for col in cols))
        rej.add('cf', rot='COSTES FIJOS (los del P&L)', fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(col, '=' + pyg.r('tcf')) for col in cols]
                    + [('E', '="Los costes fijos no cambian con el escenario: '
                        'por eso son fijos"')]))
        rej.add('rai', rot='RESULTADO ANTES DE IMPUESTOS', bold=True,
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=' + R.c('mb', c) + '-'
                           + R.c('cf', c))) for col in cols))
        rej.add('is', rot='Impuesto de Sociedades', fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(col, (lambda R, c=col: '=MAX(0,' + R.c('rai', c) + '-'
                            + P('bin_inicial') + ')*' + P('is_nueva')))
                     for col in cols]
                    + [('E', '="Al tipo de entidad de nueva creación, '
                        'compensando las bases negativas anteriores"')]))
        rej.add('neto', rot='RESULTADO NETO', bold=True, fmt=motor.FMT_EUR0,
                formulas=dict(
                    (col, (lambda R, c=col: '=' + R.c('rai', c) + '-'
                           + R.c('is', c))) for col in cols))
        rej.add('be', rot='Cubiertos/día para el equilibrio',
                fmt=motor.FMT_DEC,
                formulas=dict(
                    (col, (lambda R, c=col: '=IF(' + R.c('ticket', c) + '-'
                           + R.c('tcv', c) + '/(' + R.c('cub', c) + '*'
                           + R.c('dias', c) + ')<=0,"",' + R.c('cf', c) + '/('
                           + R.c('ticket', c) + '-' + R.c('tcv', c) + '/('
                           + R.c('cub', c) + '*' + R.c('dias', c) + '))/'
                           + R.c('dias', c) + ')')) for col in cols))
        # RD-19 — los escenarios no tenían consecuencia ni de personal ni de
        # caja: el pesimista dejaba el coste laboral en el 53 % de las ventas
        # y se comía el 78 % del fondo de maniobra, y el optimista servía 95
        # cubiertos al día con la misma plantilla. Ninguna fila lo decía.
        rej.add(rot='LO QUE CADA ESCENARIO EXIGE', bold=True)
        rej.add('r_personal', rot='Coste de personal / Ventas',
                fmt=motor.FMT_PCT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=' + pyg.r('cf_personal') + '/'
                            + R.c('ingresos', c))) for col in cols]
                    + [('E', '="La MISMA plantilla en los tres escenarios. Si '
                        'en el pesimista se dispara, el ajuste hay que '
                        'hacerlo en la hoja de Personal, no aquí"')]))
        motor.semaforo_num(rej.ws, rej.c('r_personal') + ':'
                           + rej.c('r_personal', 'D'),
                           verde_si=rej.c('r_personal') + '<='
                           + pyg.r('r_personal', 'E'),
                           rojo_si=rej.c('r_personal') + '>'
                           + pyg.r('r_personal', 'E'))
        rej.add('cub_fte', rot='Cubiertos al año por jornada equivalente',
                fmt=motor.FMT_ENT,
                formulas=dict(
                    [(col, (lambda R, c=col: '=' + R.c('cub', c) + '*'
                            + R.c('dias', c) + '/'
                            + self.rej['personal'].r('total', 'C')))
                     for col in cols]
                    + [('E', '="Cubiertos que tendría que sacar cada jornada '
                        'completa del cuadro de Personal. En el optimista '
                        'dice si hace falta contratar"')]))
        # CRIT-06 — dos «saldo de caja al cierre del año 1» distintos en el
        # mismo libro: el atajo de esta hoja (fondo de maniobra + neto +
        # amortización − principal) daba 118.312,24 € y la tesorería cierra el
        # mes 12 en 166.219,82 €, un 40 % más, porque el atajo NO ve el desfase
        # de cobro/pago, la rampa de arranque ni la compensación del IVA de la
        # inversión. La columna REALISTA —la única que reproduce el caso base
        # del P&L— pasa a leer la tesorería, que es la cifra buena. Pesimista y
        # optimista conservan el atajo: no tienen tesorería propia. La nota lo
        # dice, porque es la cifra que el guion del docx cita (§4.2).
        # ⚠️ M3 / R22-CAF-10 (2026-09-05, cierra R-13 y CRIT-06). La fila
        # seguía mezclando DOS MÉTODOS en la misma línea: la columna realista
        # leía la tesorería mensual y las otras dos un atajo. Estaba
        # declarado en el rótulo y en la nota, pero leídas EN FILA las tres
        # columnas no eran comparables: en la cafetería el optimista parecía
        # sacar 16.320 € más de caja que el realista cuando, con el mismo
        # método, la diferencia son 43.101 €. Ahora las TRES columnas usan el
        # método estimado —lo único que se puede calcular para el pesimista y
        # el optimista, que no tienen tesorería mensual propia— y el saldo
        # real de la tesorería baja a su propia fila, rotulada como lo que
        # es. La cifra buena del caso base no se pierde: está justo debajo.
        def _caja(col):
            return (lambda R, c=col: '='
                    + self.rej['inversion'].r('fondo') + '+'
                    + R.c('neto', c) + '+' + pyg.r('cf_amort')
                    + '-' + self.rej['financiacion'].r('cap_1'))

        rej.add('caja_cierre',
                rot='Saldo de caja al cierre del año 1 (estimado, el mismo '
                    'método en los tres escenarios)',
                fmt=motor.FMT_EUR0, bold=True,
                formulas=dict(
                    [(col, _caja(col)) for col in cols]
                    + [('E', '="Saldo ESTIMADO: fondo de maniobra más el '
                        'resultado del año, devolviendo la amortización (que '
                        'no se paga) y restando el principal del préstamo. No '
                        'incluye el desfase de cobros y pagos ni la '
                        'liquidación del IVA, así que las tres columnas se '
                        'pueden comparar entre sí. En rojo, el escenario se '
                        'queda sin caja"')]))
        motor.semaforo_num(rej.ws, rej.c('caja_cierre') + ':'
                           + rej.c('caja_cierre', 'D'),
                           verde_si=rej.c('caja_cierre') + '>0',
                           rojo_si=rej.c('caja_cierre') + '<=0')
        rej.add('caja_real',
                rot='Saldo real del mes 12 de la hoja de Tesorería (sólo '
                    'realista)',
                fmt=motor.FMT_EUR0,
                formulas={
                    'C': (lambda R: '=' + self.rej['tesoreria'].r(
                        'saldo', get_column_letter(2 + 11))),
                    'E': '="El saldo REAL del caso base, con su desfase de '
                         'cobros y pagos y su liquidación de IVA. Sólo existe '
                         'para el escenario realista, que es el único con '
                         'tesorería mensual propia: por eso no está en la '
                         'fila de arriba, donde se compararía contra dos '
                         'estimaciones"'})
        self.pendientes.append(rej)
        fila = rej.ultima + 2
        motor.val(ws, 'A' + str(fila),
                  'La columna «Realista» lee sus tres datos de la hoja de '
                  'Supuestos, así que reproduce EXACTAMENTE el año 1 del P&L. '
                  'Los otros dos escenarios usan las mismas tasas de coste: '
                  'lo único que cambia son los cubiertos, el ticket y los '
                  'días. Los COSTES FIJOS sí son los mismos en los tres: los '
                  'varios e imprevistos, que son un porcentaje de las ventas, '
                  'van arriba con los variables y cada escenario carga los '
                  'suyos.', wrap=True)
        ws.row_dimensions[fila].height = 44
        for i, texto in enumerate(pie):
            motor.val(ws, 'A' + str(fila + 1 + i), texto)
        motor.anchos(ws, {'A': 40, 'B': 16, 'C': 16, 'D': 16, 'E': 52})
        motor.print_setup(ws, header_row=cab)
        return rej

    def _leer_escenarios(self, ws, cab):
        """Cubiertos/ticket/días de los escenarios extremos del fichero."""
        propios = self.dato('ESCENARIOS') or {}
        fuera = {'pesimista': [None, None, None],
                 'optimista': [None, None, None]}
        patrones = ((r'^(clientes|cubiertos)', 0), (r'^ticket', 1),
                    (r'd[ií]as', 2))
        # ⚠️ RT-02 — el bucle se quedaba con la ÚLTIMA fila que casaba, y la
        # hoja que este mismo grupo escribe tiene DOS filas que empiezan por
        # «Cubiertos/día»: el driver (input) y «Cubiertos/día para el
        # equilibrio» (fórmula). En la 2.ª pasada leía la fórmula, `_num`
        # devolvía None y los tres drivers de los escenarios se recalculaban
        # desde el caso base: idempotencia rota en el molde A-β, que es el de
        # cuatro de los diez productos. Se toma la PRIMERA fila que casa y
        # sólo si trae un número, no una fórmula.
        for r in range(cab + 1, ws.max_row + 1):
            rot = motor._rotulo_de_fila(ws, r, max_col=1)
            if not rot:
                continue
            for patron, idx in patrones:
                if fuera['pesimista'][idx] is not None:
                    continue
                if not re.search(patron, rot, re.I):
                    continue
                b, d = ws['B' + str(r)].value, ws['D' + str(r)].value
                if motor._es_formula(b) or motor._es_formula(d):
                    continue
                fuera['pesimista'][idx] = _num(b)
                fuera['optimista'][idx] = _num(d)
        base = [self.p.valor('cubiertos_dia'), self.p.valor('ticket_medio'),
                self.p.valor('dias_apertura')]
        for clave, factor in (('pesimista', 0.72), ('optimista', 1.25)):
            if clave in propios:
                fuera[clave] = list(propios[clave])
            for i in range(3):
                if fuera[clave][i] is None:
                    ref_ = _num(base[i], 0) or 0
                    fuera[clave][i] = round(ref_ * (factor if i < 2 else 1), 2)
        return fuera

    # -- §2.7 -------------------------------------------------------------
    def tesoreria(self):
        """`Tesorería 12 meses`: en qué mes se agota la caja (§2.7)."""
        ws = self.ws_tesoreria
        _limpiar_area(ws, 1, max(ws.max_row, 70), 15)
        motor.val(ws, 'A1', 'PREVISIÓN DE TESORERÍA — AÑO 1', bold=True)
        motor.val(ws, 'A2', 'El P&L dice si el negocio gana dinero; esta hoja '
                  'dice si le queda caja para llegar a fin de mes. Es la '
                  'primera que mira un banco.', wrap=True)
        cab = 4
        motor.val(ws, 'A' + str(cab), 'Concepto', bold=True)
        # el «(€)» de la cabecera no es adorno: sin él, `motor` lee «Mes» como
        # recuento y borra el formato de euro de las doce columnas
        meses = tuple('Mes ' + str(i) + ' (€)' for i in range(1, 13))
        cols = [get_column_letter(2 + i) for i in range(12)]
        for i, nombre in enumerate(meses):
            motor.val(ws, cols[i] + str(cab), nombre, bold=True)
        motor.val(ws, 'N' + str(cab), 'Año (€)', bold=True)
        pyg_ = self.rej['pyg']
        rej = Rejilla(ws, cab + 1,
                      driver=lambda R, c, k: pyg_.r('ingresos'),
                      cols_driver=tuple(get_column_letter(2 + i)
                                        for i in range(12)) + ('B', 'N'))
        self.rej['tesoreria'] = rej
        pyg = self.rej['pyg']
        inv = self.rej['inversion']
        fin = self.rej['financiacion']
        P = self.p.ref
        estacion = list(self.dato('ESTACIONALIDAD') or ([1.0 / 12] * 12))

        def suma_ano(clave):
            return (lambda R, k=clave: '=SUM(' + R.c(k, cols[0]) + ':'
                    + R.c(k, cols[11]) + ')')

        def por_meses(clave, rot, ref_anual, nota=None, signo='', fmt=None,
                      reparto='peso'):
            """Reparte un importe anual entre los doce meses.

            `reparto='peso'` sigue la curva de actividad (ventas y compras) y
            `'lineal'` divide entre doce: el alquiler y la cuota del préstamo
            NO bajan en agosto, y repartirlos por actividad escondía justo el
            mes en el que la caja se tensa.
            """
            factor = ('*' + '{peso}') if reparto == 'peso' else '/12'
            formulas = dict(
                (cols[i], (lambda R, c=cols[i], k=clave:
                           '=' + signo + ref_anual
                           + factor.replace('{peso}', R.c('peso', c))))
                for i in range(12))
            formulas['N'] = suma_ano(clave)
            rej.add(clave, rot=rot, fmt=fmt or motor.FMT_EUR0,
                    formulas=formulas, valores={'O': nota} if nota else None)

        # ---- curva de actividad: estacionalidad × rampa (RD-17 / RC-03) ---
        # La fila que había era una ESTACIONALIDAD (mínimo 6,2 % en agosto,
        # máximo 10,6 % en diciembre), no una curva de arranque: el mes 1
        # facturaba el 7,0 % del año con los cubiertos de crucero, así que el
        # saldo mínimo era el del primer mes y la caja NO PODÍA agotarse
        # nunca. La hoja que el propio libro presenta como «la primera que
        # mira un banco» estaba construida para no poder dar un susto.
        rampa1 = P('rampa_mes1')
        meses_rampa = P('meses_rampa')
        rej.add('estacional', rot='Estacionalidad del mes (suma 100 %)',
                fmt=motor.FMT_PCT, verdes=tuple(cols),
                valores=dict((cols[i], round(estacion[i], 4))
                             for i in range(12)),
                formulas={'N': suma_ano('estacional'),
                          'O': 'Perfil de tu zona: agosto flojo en ciudad y '
                               'fuerte en costa. Tiene que sumar 100 %'})
        # T7/panadería (2026-08-29): la igualdad ESTRICTA `=1`/`<>1` es
        # inherentemente frágil sobre un `SUM()` de 12 decimales en coma
        # flotante — probado en Python con los pesos de ESTE módulo, que
        # suman 1,000 en aritmética decimal exacta: `sum(...)` da
        # `0.9999999999999999`, no `1.0`. No es un caso raro: CUALQUIER
        # conjunto de 12 pesos con dos o tres decimales tropieza con el
        # mismo redondeo binario (probado también con `[1/12]*12` y con
        # `[0.1]*10`). El REF-03 de esta misma familia («…tapas-bar-ref.
        # json») pedía «que la suma dé exactamente 1,000», pero ningún
        # conjunto de decimales lo hace de forma fiable en IEEE754: la regla
        # buena es la ROJA de verdad (1,02, un 2 % de más) contra el RUIDO
        # de redondeo (10⁻¹⁶), no una igualdad exacta entre los dos.
        motor.semaforo_num(ws, rej.c('estacional', 'N') + ':'
                           + rej.c('estacional', 'N'),
                           verde_si='ABS(' + rej.c('estacional', 'N')
                           + '-1)<0.005',
                           rojo_si='ABS(' + rej.c('estacional', 'N')
                           + '-1)>=0.005')
        rej.add('rampa', rot='Rampa de arranque (% de la actividad de '
                'crucero)', fmt=motor.FMT_PCT,
                formulas=dict(
                    [(cols[i], '=MIN(1,' + rampa1 + '+(1-' + rampa1 + ')*'
                      + str(i) + '/MAX(1,' + meses_rampa + '-1))')
                     for i in range(12)]
                    + [('O', 'Sube en línea recta desde el porcentaje del mes '
                        '1 hasta el 100 % en los meses que digas en '
                        'Supuestos. Con el mes 1 al 100 % la rampa '
                        'desaparece')]))
        rej.add('peso', rot='Reparto de la actividad por mes (calculado)',
                fmt=motor.FMT_PCT,
                formulas=dict(
                    [(cols[i], (lambda R, c=cols[i]:
                                '=' + R.c('estacional', c) + '*'
                                + R.c('rampa', c) + '/SUMPRODUCT('
                                + R.c('estacional', cols[0], absoluta=True)
                                + ':'
                                + R.c('estacional', cols[11], absoluta=True)
                                + ',' + R.c('rampa', cols[0], absoluta=True)
                                + ':' + R.c('rampa', cols[11], absoluta=True)
                                + ')')) for i in range(12)]
                    + [('N', suma_ano('peso')),
                       ('O', 'Estacionalidad × rampa, normalizado a 100 %: el '
                        'AÑO factura lo que dice el P&L, pero repartido como '
                        'lo reparte un local que acaba de abrir')]))
        rej.add(rot='COBROS', bold=True)
        # RD-35 — el modelo cobraba el 100 % de las ventas del mes en el mes,
        # sin ninguna celda de desfase, cuando §2.7 lo pide expresamente. En
        # un bar casi todo es contado, pero la tarjeta y las plataformas de
        # reserva y de reparto liquidan a D+1/D+30: con 0 días el resultado
        # no cambia y el modelo queda demostrable.
        d_cobro = P('dias_cobro')
        # A3 — el IVA que se cobra al cliente sale de la tabla de líneas de
        # venta del P&L (pesos × tipos), no de dos porcentajes globales.
        ventas_iva = ('(' + pyg.r('ingresos') + '+' + pyg.r('ingresos') + '*'
                      + self.iva_ventas + ')')
        cobros = {}
        for i in range(12):
            actual = ('%s*%s*(1-MIN(30,%s)/30)'
                      % (ventas_iva, '{p}', d_cobro))
            actual = actual.replace('{p}', '{peso_i}')
            if i == 0:
                formula = ('=' + ventas_iva + '*{peso_i}*(1-MIN(30,'
                           + d_cobro + ')/30)')
            else:
                formula = ('=' + ventas_iva + '*({peso_i}*(1-MIN(30,'
                           + d_cobro + ')/30)+{peso_p}*MIN(30,' + d_cobro
                           + ')/30)')
            cobros[cols[i]] = (
                lambda R, f=formula, c=cols[i], p=(cols[i - 1] if i else None):
                f.replace('{peso_i}', R.c('peso', c)).replace(
                    '{peso_p}', R.c('peso', p) if p else '0'))
        cobros['N'] = suma_ano('cobros')
        cobros['O'] = ('Ventas del mes con el IVA repercutido, corregidas por '
                       'los días medios de cobro de la hoja de Supuestos')
        rej.add('cobros', rot='Ventas cobradas (con IVA repercutido)',
                fmt=motor.FMT_EUR0, formulas=cobros)
        rej.add(rot='PAGOS', bold=True)
        # RD-13 — cada compra a SU tipo, leído de la columna de IVA del P&L,
        # y con el desfase de pago a proveedor de Supuestos (RD-35).
        d_pago = P('dias_pago')
        compras = ('(' + pyg.r('cv_comida') + '*(1+' + pyg.r('cv_comida', 'G')
                   + ')+' + pyg.r('cv_bebida') + '*(1+'
                   + pyg.r('cv_bebida', 'G') + ')+' + pyg.r('cv_cons')
                   + '*(1+' + pyg.r('cv_cons', 'G') + ')+'
                   + pyg.r('cv_deliv') + '*(1+' + pyg.r('cv_deliv', 'G')
                   + ')+' + pyg.r('cv_tpv') + '*(1+' + pyg.r('cv_tpv', 'G')
                   + ')+' + pyg.r('cv_varios') + '*(1+'
                   + pyg.r('cv_varios', 'G') + '))')
        pagos = {}
        for i in range(12):
            if i == 0:
                formula = ('=-' + compras + '*{peso_i}*(1-MIN(30,' + d_pago
                           + ')/30)')
            else:
                formula = ('=-' + compras + '*({peso_i}*(1-MIN(30,' + d_pago
                           + ')/30)+{peso_p}*MIN(30,' + d_pago + ')/30)')
            pagos[cols[i]] = (
                lambda R, f=formula, c=cols[i], p=(cols[i - 1] if i else None):
                f.replace('{peso_i}', R.c('peso', c)).replace(
                    '{peso_p}', R.c('peso', p) if p else '0'))
        pagos['N'] = suma_ano('p_var')
        # B13 / R-24 — con días de pago a proveedor, la compra del último mes
        # se paga en enero: la columna «Año» de ESTA fila es caja de doce
        # meses, no el coste del ejercicio, y el lector la compara con el P&L.
        # ⚠️ M20 / REF22-BAR-02 (2026-09-05): la nota decía «DOCE
        # mensualidades» y son ONCE — enero no paga nada, porque el mes
        # anterior no se compró. Es justo el número que hace falta para
        # cuadrar los −141.152 € de la fila contra los 160.067 € de coste
        # anual; con «doce» el desfase quedaba inexplicado. Y se generaliza:
        # con B59 días de pago, las mensualidades recogidas son
        # 12 − REDONDEO.MÁS(B59/30).
        pagos['O'] = ('="Materia prima, consumibles, comisiones e '
                      'imprevistos, cada uno con SU tipo de IVA (la columna '
                      'del P&L) y con los días de pago a proveedor de '
                      'Supuestos."&IF(' + d_pago + '=0,""," Con "&TEXT('
                      + d_pago + ',"0")&" días de pago, las primeras compras '
                      'se pagan ya en el año siguiente: la columna «Año» de '
                      'esta fila recoge "&TEXT(MAX(0,12-ROUNDUP('
                      + d_pago + '/30,0)),"0")&" mensualidades de compra, no '
                      'el coste del ejercicio.")')
        rej.add('p_var', rot='Compras y costes variables (IVA incluido)',
                fmt=motor.FMT_EUR0, formulas=pagos)
        fijos_caja = ('(' + pyg.r('tcf') + '-' + pyg.r('cf_personal') + '-'
                      + pyg.r('cf_amort') + '-' + pyg.r('cf_int') + '+'
                      + pyg.r('iva_fij') + ')')
        por_meses('p_fijos', 'Costes fijos de explotación (IVA incluido)',
                  fijos_caja, signo='-', reparto='lineal',
                  nota='Los fijos que salen de caja, cada uno con su tipo de '
                       'IVA: los seguros están exentos y no llevan')
        # RD-18 / RT-27 — las nóminas se repartían en doce partes iguales
        # pese a que el modelo trabaja con 14 pagas: las dos extras no
        # aparecían en su mes, que es justo cuando revienta la tesorería de
        # un negocio estacional.
        # el mismo suelo de 12 mensualidades que la hoja de Personal
        # (RT-08): con un 0 el reparto se iría a infinito
        pagas = 'MAX(12,' + P('pagas') + ')'
        base_mes = pyg.r('cf_personal') + '/' + pagas
        extra = '(' + pagas + '-12)/2'
        nominas = dict(
            (cols[i], (lambda R, c=cols[i], n=i + 1:
                       '=-' + base_mes + '*(1+IF(' + str(n) + '='
                       + P('mes_paga_1') + ',' + extra + ',0)+IF(' + str(n)
                       + '=' + P('mes_paga_2') + ',' + extra + ',0))'))
            for i in range(12))
        nominas['N'] = suma_ano('p_personal')
        nominas['O'] = ('Doce mensualidades más las pagas extra en los meses '
                        'que digas en Supuestos. La suma del año es la misma '
                        'que la del P&L')
        rej.add('p_personal', rot='Nóminas y Seguridad Social',
                fmt=motor.FMT_EUR0, formulas=nominas)
        por_meses('p_int', 'Intereses del préstamo', fin.r('int_1'),
                  signo='-', reparto='lineal')
        por_meses('p_principal', 'Devolución de principal del préstamo',
                  fin.r('cap_1'), signo='-', reparto='lineal')
        # ---- IVA: memoria, compensación y liquidación trimestral ---------
        rej.add('iva_rep', rot='IVA repercutido del mes (memoria)',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(cols[i], (lambda R, c=cols[i]: '=' + pyg.r('ingresos')
                                + '*' + R.c('peso', c) + '*'
                                + self.iva_ventas)) for i in range(12)]
                    + [('N', suma_ano('iva_rep')),
                       ('O', 'Cada línea de venta a SU tipo, leído de la '
                        'columna de IVA del P&L: en sala, alcohol incluido, '
                        'el reducido (art. 91.Uno.2.2.º de la Ley 37/1992 '
                        'del IVA); el alcohol que sale por delivery, al '
                        'general; el pan común y las harinas panificables, '
                        'al superreducido (art. 91.Dos.1.1.º)')]))
        rej.add('iva_sop', rot='IVA soportado del mes (memoria)',
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(cols[i], (lambda R, c=cols[i]:
                                '=' + pyg.r('iva_var') + '*' + R.c('peso', c)
                                + '+' + pyg.r('iva_fij') + '/12'))
                      for i in range(12)]
                    + [('N', suma_ano('iva_sop')),
                       ('O', 'Cada partida a su tipo, leído de la columna de '
                        'IVA del P&L: los seguros y las comisiones de los '
                        'medios de pago están exentos')]))
        # RD-11 / RT-16 — el IVA de la inversión (19.542,60 € en el caso base
        # de partida) se financiaba a siete años y NO se recuperaba en
        # ninguna parte del modelo, mientras el libro afirmaba dos veces que
        # es recuperable por el modelo 303. Y la liquidación estaba truncada
        # con `-MAX(0,…)`: cuando el soportado supera al repercutido —lo
        # normal en el arranque— la hoja escribía 0 en vez de arrastrar la
        # compensación al trimestre siguiente.
        trimestres = (4, 7, 10)
        pend = {}
        for i in range(12):
            if i + 1 == trimestres[0]:
                pend[cols[i]] = '=' + inv.r('iva')
            elif i + 1 in trimestres:
                anterior = cols[i - 3]
                pend[cols[i]] = (lambda R, a=anterior:
                                 '=MAX(0,-' + R.c('iva_cuota', a) + ')')
        pend['O'] = ('El IVA soportado de la INVERSIÓN entra aquí: se '
                     'compensa contra el repercutido de los trimestres '
                     'siguientes hasta agotarlo (modelo 303)')
        rej.add('iva_pend', rot='IVA a compensar arrastrado (inversión '
                'incluida)', fmt=motor.FMT_EUR0, formulas=pend)
        cuota = {}
        for i in range(12):
            if i + 1 in trimestres:
                a, b = i - 3, i - 1
                cuota[cols[i]] = (
                    lambda R, a=a, b=b, c=cols[i]:
                    '=SUM(' + R.c('iva_rep', cols[a]) + ':'
                    + R.c('iva_rep', cols[b]) + ')-SUM('
                    + R.c('iva_sop', cols[a]) + ':'
                    + R.c('iva_sop', cols[b]) + ')-' + R.c('iva_pend', c))
        cuota['O'] = ('Repercutido menos soportado del trimestre menos lo que '
                      'venía arrastrado. En negativo NO se paga: se compensa '
                      'en el trimestre siguiente')
        rej.add('iva_cuota', rot='Resultado de la liquidación trimestral',
                fmt=motor.FMT_EUR0, formulas=cuota)
        liq = dict(
            (cols[i], (lambda R, c=cols[i]: '=IF(' + R.c('iva_cuota', c)
                       + '<=0,"",-' + R.c('iva_cuota', c) + ')'))
            for i in range(12) if i + 1 in trimestres)
        liq['N'] = suma_ano('iva_liq')
        # M20 / REF22-BAR-07 — la regla de formato condicional de esta fila
        # compara el pago del 303 contra el saldo POSTERIOR al pago, así que
        # es una alerta de liquidez más temprana que «te quedas en números
        # rojos». No se cambia el criterio (avisar antes es lo útil): se
        # DICE, que era lo que faltaba.
        liq['O'] = ('Sólo sale caja cuando el resultado del trimestre es '
                    'positivo. Los meses sin liquidación van en blanco, no a '
                    'cero. La celda se pinta en ROJO cuando el pago del 303 '
                    'se lleva más caja de la que queda ese mes')
        rej.add('iva_liq', rot='Pago del IVA (modelo 303)',
                fmt=motor.FMT_EUR0, formulas=liq)
        # CRIT-01 — el sumando del IVA va GUARDADO. `iva_liq` vale `""` en los
        # meses sin liquidación (es su diseño: «los meses sin liquidación van
        # en blanco, no a cero»), y en Excel `número + ""` es #¡VALOR!, que el
        # `IFERROR` de la envoltura convertía en `""`. Resultado medido: los
        # meses 4, 7 y 10 del FLUJO en blanco y, por cascada, NUEVE de los doce
        # SALDOS ACUMULADOS —la fila que el propio libro anuncia como «la
        # primera que mira un banco»— con el flujo de caja libre del año 1 un
        # 23,8 % por debajo del real y el payback en 2,98 años en vez de 2,58.
        # Ningún gate lo cazaba: `data_only` clasificaba las 12 celdas como
        # «vacías por diseño». `SUM` sí ignora el texto; la suma con `+` no.
        rej.add('flujo', rot='FLUJO DEL MES', bold=True, fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(cols[i], (lambda R, c=cols[i]: '=SUM(' + R.c('cobros', c)
                                + ':' + R.c('p_principal', c) + ')+IF('
                                + R.c('iva_liq', c) + '="",0,'
                                + R.c('iva_liq', c) + ')'))
                     for i in range(12)]
                    + [('N', suma_ano('flujo'))]))
        rej.add('saldo', rot='SALDO ACUMULADO DE CAJA', bold=True,
                fmt=motor.FMT_EUR0,
                formulas=dict(
                    [(cols[0], (lambda R: '=' + inv.r('fondo') + '+'
                                + R.c('flujo', cols[0])))]
                    + [(cols[i], (lambda R, c=cols[i], p=cols[i - 1]:
                                  '=' + R.c('saldo', p) + '+'
                                  + R.c('flujo', c))) for i in range(1, 12)]))
        motor.semaforo_num(ws, rej.c('saldo', cols[0]) + ':'
                           + rej.c('saldo', cols[11]),
                           verde_si=rej.c('saldo', cols[0]) + '>0',
                           rojo_si=rej.c('saldo', cols[0]) + '<0')
        # B9 / REF-09 — el semáforo genérico del motor pintaba de ROJO la
        # fila «Resultado de la liquidación trimestral» en los trimestres
        # NEGATIVOS, que son justo aquellos en los que no hay que pagar IVA
        # (se compensa en el siguiente, y la nota de la fila ya lo dice). El
        # motor la excluye ahora por rótulo y el rojo se reserva para el caso
        # que sí duele: que el pago del 303 se coma el saldo de caja del mes.
        motor.semaforo_num(ws, rej.c('iva_liq', cols[0]) + ':'
                           + rej.c('iva_liq', cols[11]),
                           rojo_si='-' + rej.c('iva_liq', cols[0]) + '>'
                           + rej.c('saldo', cols[0]))
        # `MIN` sobre doce celdas vacías devuelve 0, y un 0 aquí pinta el
        # semáforo de VERDE en un libro sin datos: hay que preguntar antes si
        # hay algún número que comparar
        rej.add('minimo', rot='Saldo mínimo del año', fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=IF(COUNT(' + R.c('saldo', cols[0])
                                + ':' + R.c('saldo', cols[11]) + ')=0,"",MIN('
                                + R.c('saldo', cols[0]) + ':'
                                + R.c('saldo', cols[11]) + '))'),
                          'D': '="Si es negativo, sube el fondo de maniobra o '
                               'el préstamo: te quedas sin caja"'})
        motor.semaforo_num(ws, rej.c('minimo') + ':' + rej.c('minimo'),
                           verde_si=rej.c('minimo') + '>=0',
                           rojo_si=rej.c('minimo') + '<0')
        rej.add('mes_minimo', rot='Mes en el que la caja toca fondo',
                fmt=motor.FMT_ENT,
                formulas={'B': (lambda R: '=IF(COUNT('
                                + R.c('saldo', cols[0]) + ':'
                                + R.c('saldo', cols[11]) + ')=0,"",MATCH(MIN('
                                + R.c('saldo', cols[0]) + ':'
                                + R.c('saldo', cols[11]) + '),'
                                + R.c('saldo', cols[0]) + ':'
                                + R.c('saldo', cols[11]) + ',0))'),
                          'D': '="Es la respuesta a la pregunta que decide '
                               'una operación bancaria: en qué mes se agota '
                               'la caja"'})
        # ---- payback ÚNICO del proyecto (DOM-17, RD-06, RT-22) -----------
        rej.add(rot='RETORNO DE LA INVERSIÓN', bold=True)
        # RT-22 — la hoja contenía DOS flujos de caja del año 1 que no
        # coincidían (63.068 € en la tabla mensual y 49.014 € en la fila del
        # payback) y el payback usaba el que la hoja NO había construido. El
        # año 1 pasa a ser el de la propia tabla; los años 2 y 3, que no
        # tienen tabla mensual, se estiman y la nota lo dice.
        # ⚠️ A1 / R-06 / REF-07-motor / M-08 (2026-09-05). El payback CONTABA
        # LA DEUDA DOS VECES: el numerador era la necesidad TOTAL de caja
        # (CAPEX + fondo + IVA soportado, con el préstamo dentro) y el
        # denominador, un flujo del que ya se habían restado los intereses Y
        # la devolución de principal. Salía «Más de 3 años» en cuatro de los
        # cinco hermanos, contradiciendo la referencia «Retorno de la
        # inversión: 24-36 meses» que el propio libro publica. Ahora es el
        # payback DEL PROYECTO: inversión sin el IVA (que se recupera por el
        # 303) entre el flujo de caja libre ANTES del servicio de la deuda.
        # Si el negocio puede pagar la cuota lo dice el DSCR, no el payback.
        # ⚠️ M2 / R22-CAF-09 / REF22-BAR-01 / R22-PAN-04 (2026-09-05). El
        # payback publicado MEZCLABA DOS BASES: el año 1 iba con la caja REAL
        # de la tabla mensual —inflada por el float del IVA (el crédito de la
        # inversión hace que el año 1 no pague un euro por el 303) y por los
        # 30 días de crédito del proveedor— y los años 2 y 3 con la base
        # contable. Medio año de payback dependía de esa elección, siempre a
        # favor: en el bar salía 1,9 cuando NINGUNA lectura coherente lo daba
        # (2,0 / 2,1 / 2,3 según cómo se cerrase el círculo). Y el numerador
        # ya venía limpio de IVA («se recupera por el 303»), así que el mismo
        # IVA se contaba dos veces. Ahora las TRES anualidades van en base
        # contable —resultado neto + amortización + intereses— y el flujo
        # real del año 1 se queda como fila INFORMATIVA, que es lo que
        # siempre debió ser.
        rej.add('fcf_1',
                rot='Flujo de caja libre antes de la deuda, año 1',
                fmt=motor.FMT_EUR0,
                formulas={'B': '=' + pyg.r('neto') + '+' + pyg.r('cf_amort')
                          + '+' + fin.r('int_1'),
                          'D': '="Resultado neto más amortización (que no se '
                               'paga) más los intereses de ese año: el '
                               'principal no se resta porque este flujo es '
                               'ANTES de la deuda. La MISMA base que los años '
                               '2 y 3, para que el payback compare peras con '
                               'peras"'})
        for i, clave in ((2, 'fcf_2'), (3, 'fcf_3')):
            col = ('B', 'C', 'D')[i - 1]
            rej.add(clave, rot='Flujo de caja libre antes de la deuda, año '
                    + str(i), fmt=motor.FMT_EUR0,
                    formulas={'B': '=' + pyg.r('neto', col) + '+'
                              + pyg.r('cf_amort', col) + '+'
                              + fin.r('int_%d' % i),
                              'D': '="Resultado neto más amortización (que no '
                                   'se paga) más los intereses de ese año: el '
                                   'principal no se resta porque este flujo '
                                   'es ANTES de la deuda. Sin tabla mensual, '
                                   'es una estimación"'})
        rej.add('fcf_1_est',
                rot='Informativo: flujo REAL de caja del año 1 (tabla '
                    'mensual, antes de la deuda)',
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=' + R.c('flujo', 'N') + '+'
                                + fin.r('int_1') + '+' + fin.r('cap_1')),
                          'D': '="La suma de los doce FLUJO DEL MES de '
                               'arriba, devolviéndole los intereses y el '
                               'principal que se pagaron ese año. NO es la '
                               'que usa el payback: lleva dentro el IVA que '
                               'todavía no se ha liquidado y el crédito del '
                               'proveedor, dos cosas que hay que devolver y '
                               'que los años 2 y 3 no tienen"'})
        rej.add('inv_recup',
                rot='Inversión a recuperar (sin el IVA, que se recupera por '
                    'el 303)',
                fmt=motor.FMT_EUR0,
                formulas={
                    'B': '=' + inv.r('caja') + '-' + inv.r('iva'),
                    'D': '="CAPEX más fondo de maniobra: la necesidad total '
                         'de caja MENOS el IVA soportado de la inversión, que '
                         'no se pierde (se recupera con el modelo 303 y esta '
                         'misma hoja lo compensa trimestre a trimestre). El '
                         'fondo de maniobra sí se recupera al cerrar, pero '
                         'hasta entonces está inmovilizado"'})
        def _cascada(num):
            """Payback a tres años sobre `num` (referencia o clave del num)."""
            return (lambda R, n=num: '=IF(' + R.c('fcf_1') + '>=' + n
                    + ',ROUND(' + n + '/' + R.c('fcf_1') + ',1),IF('
                    + R.c('fcf_1') + '+' + R.c('fcf_2') + '>=' + n
                    + ',ROUND(1+(' + n + '-' + R.c('fcf_1') + ')/'
                    + R.c('fcf_2') + ',1),IF(' + R.c('fcf_1') + '+'
                    + R.c('fcf_2') + '+' + R.c('fcf_3') + '>=' + n
                    + ',ROUND(2+(' + n + '-' + R.c('fcf_1') + '-'
                    + R.c('fcf_2') + ')/' + R.c('fcf_3') + ',1),'
                    + '"Más de 3 años")))')

        rej.add('payback',
                rot='Payback del proyecto (años), antes de la deuda',
                # A1/B4 — la celda puede devolver el TEXTO «Más de 3 años»:
                # con un formato numérico encima, cualquier suma del usuario
                # da #¡VALOR!. Va en General, y el número se redondea dentro
                # de la fórmula para que se lea igual de bien.
                fmt='General', bold=True,
                formulas={
                    'B': (lambda R: _cascada(R.c('inv_recup'))(R)),
                    'D': '="Mide cuánto tarda el negocio en devolver la '
                         'inversión con lo que genera ANTES de pagar al '
                         'banco, con la MISMA base de flujo en los tres '
                         'años; si puedes pagar la cuota lo dice el DSCR de '
                         'la hoja de Financiación. El Word cita esta celda, '
                         'no recalcula"'})
        # M2 — el fondo de maniobra NO se consume: se recupera al cerrar, y
        # mientras tanto alarga el plazo de un proyecto que sobre el activo
        # que se compra ya está devuelto. Va como fila informativa, no como
        # la cifra que se publica.
        rej.add('payback_capex',
                rot='Payback sobre el CAPEX, sin fondo de maniobra (años)',
                fmt='General',
                formulas={
                    'B': (lambda R: _cascada(inv.r('capex'))(R)),
                    'D': '="El mismo cálculo sobre lo que se COMPRA (obra, '
                         'equipamiento, lanzamiento y constitución), sin el '
                         'fondo de maniobra: ese dinero no se gasta, se '
                         'inmoviliza y vuelve al cerrar el negocio. '
                         'Informativo: el payback que se publica es el de '
                         'arriba"'})
        self.pendientes.append(rej)
        fila = rej.ultima + 2
        motor.val(ws, 'A' + str(fila),
                  'El IVA del cuarto trimestre se liquida en enero del año '
                  'siguiente, así que no aparece en este cuadro. El Impuesto '
                  'de Sociedades del año 1 se paga en julio del año 2. La '
                  'cuota del préstamo se reparte en doce partes iguales '
                  'siguiendo el cuadro ANUAL de la hoja de Financiación, y '
                  'las nóminas en doce mensualidades más las dos pagas extra '
                  'en su mes.', wrap=True)
        motor.anchos(ws, dict([('A', 46)] + [(c, 12) for c in cols]
                              + [('N', 14), ('O', 60)]))
        motor.print_setup(ws, header_row=cab, landscape=True,
                          congelar='B' + str(cab + 1))
        return rej


    # -- §2.8 -------------------------------------------------------------
    # -- §2.8 -------------------------------------------------------------
    def financiacion(self):
        """`Financiación`: usos y orígenes + cuadro francés (§2.8)."""
        ws = self.ws_financiacion
        _limpiar_area(ws, 1, max(ws.max_row, 80), 8)
        motor.val(ws, 'A1', 'PLAN DE FINANCIACIÓN', bold=True)
        motor.val(ws, 'A2', 'Qué hace falta, de dónde sale y cuánto cuesta '
                  'devolverlo. Es la hoja que el banco pide junto con el P&L.',
                  wrap=True)
        cab = 4
        for i, texto in enumerate(('Concepto', 'Importe', 'Notas')):
            motor.val(ws, get_column_letter(i + 1) + str(cab), texto,
                      bold=True)
        rej = Rejilla(ws, cab + 1)
        self.rej['financiacion'] = rej
        P = self.p.ref
        pyg = self.rej['pyg']
        rej.add(rot='ORIGEN DE FONDOS', bold=True)
        rej.add('propios', rot='Recursos propios de los socios',
                fmt=motor.FMT_EUR0, formulas={'B': '=' + P('recursos_propios')})
        rej.add('prestamo', rot='Préstamo bancario', fmt=motor.FMT_EUR0,
                formulas={'B': '=' + P('prestamo')})
        for clave, rot, nota in (
                ('ico', 'Línea ICO (avalada por el ICO, la concede tu banco)',
                 'Consulta las líneas del ejercicio en curso en ico.es'),
                ('enisa', 'Préstamo participativo ENISA',
                 'Para sociedades; sin garantías personales'),
                ('angeles', 'Business angels o socios inversores',
                 'Entra en el capital: diluye, no se devuelve'),
                ('subvencion', 'Subvenciones autonómicas o locales',
                 'Suelen cobrarse DESPUÉS de justificar el gasto')):
            rej.add(clave, rot=rot, fmt=motor.FMT_EUR0, verdes=('B',),
                    valores={'B': 0, 'C': nota})
        rej.add('origen', rot='TOTAL ORIGEN DE FONDOS', bold=True,
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=IF(COUNT(' + R.c('propios') + ':'
                                + R.c('subvencion') + ')=0,"",SUM('
                                + R.c('propios') + ':' + R.c('subvencion')
                                + '))')})
        rej.add(rot='USOS', bold=True)
        rej.add('usos', rot='Necesidad total de caja al arranque',
                fmt=motor.FMT_EUR0,
                formulas={'B': '=' + self.rej['inversion'].r('caja'),
                          'C': '="Inversión más el IVA que hay que adelantar"'})
        # RD-34 — los orígenes no cuadraban con los usos y el semáforo sólo
        # comprobaba que la diferencia no fuese NEGATIVA: un exceso de 20.000 €
        # pasaba igual de verde. Ahora se dice el signo, el porcentaje, y hay
        # una celda que calcula el préstamo que ajusta el origen a la
        # necesidad.
        rej.add('dif', rot='Diferencia (origen - usos)', bold=True,
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=' + R.c('origen') + '-'
                                + R.c('usos')),
                          'C': '="En rojo (negativa) el plan NO está '
                               'financiado. Muy por encima de cero tampoco es '
                               'gratis: se pagan intereses por un dinero que '
                               'no se usa"'})
        rej.add('dif_pct', rot='Diferencia sobre los usos (%)',
                fmt=motor.FMT_PCT,
                formulas={'B': (lambda R: '=' + R.c('dif') + '/'
                                + R.c('usos')),
                          'C': '="Ámbar por encima del margen que fijes: es '
                               'deuda que pagas sin necesitarla"'})
        rej.add('holgura_max', rot='Exceso de financiación admisible (%)',
                fmt=motor.FMT_PCT, verdes=('B',),
                valores={'B': 0.05,
                         'C': 'Un pequeño colchón sobre la necesidad es sano; '
                              'un 20 % de más son intereses tirados'})
        motor.semaforo_num(ws, rej.c('dif_pct') + ':' + rej.c('dif_pct'),
                           verde_si=rej.c('dif_pct') + '>=0',
                           ambar_si=rej.c('dif_pct') + '>'
                           + rej.c('holgura_max', absoluta=True),
                           rojo_si=rej.c('dif_pct') + '<0')
        rej.add('prestamo_ajustado',
                rot='Préstamo que ajustaría el origen a la necesidad',
                fmt=motor.FMT_EUR0,
                formulas={'B': (lambda R: '=MAX(0,' + R.c('usos') + '-'
                                + R.c('origen') + '+' + R.c('prestamo') + ')'),
                          'C': '="Cópialo en «Préstamo bancario solicitado» de '
                               'la hoja de Supuestos y la diferencia se pone '
                               'a cero"'})
        rej.add(rot='CONDICIONES DEL PRÉSTAMO', bold=True)
        # sin préstamo declarado el cuadro entero se apaga: un «0 €» de
        # capital pendiente al vencimiento pintaría de verde un libro vacío
        rej.add('importe', rot='Importe del principal', fmt=motor.FMT_EUR0,
                formulas={'B': '=IF(' + P('prestamo') + '=0,"",'
                          + P('prestamo') + ')'})
        rej.add('tipo', rot='Tipo de interés nominal anual',
                fmt=motor.FMT_PCT, formulas={'B': '=' + P('tipo_prestamo')})
        rej.add('plazo', rot='Plazo total (años)', fmt=motor.FMT_ENT,
                formulas={'B': '=' + P('plazo_prestamo')})
        # B1 / R-08 — con el libro vaciado, `IF("">="",0,"")` da 0: la hoja
        # publicaba una carencia de «0 años» sobre un préstamo que no existe.
        rej.add('carencia', rot='Carencia de principal aplicada (años)',
                fmt=motor.FMT_ENT,
                formulas={'B': '=IF(OR(' + P('carencia_prestamo') + '="",'
                          + P('plazo_prestamo') + '=""),"",IF('
                          + P('carencia_prestamo') + '>='
                          + P('plazo_prestamo') + ',0,'
                          + P('carencia_prestamo') + '))',
                          'C': '="Una carencia igual o mayor que el plazo no '
                               'existe: la hoja la anula en origen"'})
        # RT-06 — con el tipo al 0 % —un préstamo familiar, un ENISA o un ICO
        # bonificado, y un valor que la validación acepta— la anualidad
        # algebraica divide entre cero: el cuadro entero se vaciaba y la celda
        # «Capital pendiente al vencimiento» certificaba EN VERDE que se
        # habían devuelto 110.000 € que nadie había devuelto.
        rej.add('cuota', rot='Cuota anual durante la amortización',
                fmt=motor.FMT_EUR, bold=True,
                formulas={'B': (lambda R: '=IF(' + R.c('plazo') + '-'
                                + R.c('carencia') + '<=0,"",IF('
                                + R.c('tipo') + '=0,' + R.c('importe') + '/('
                                + R.c('plazo') + '-' + R.c('carencia') + '),'
                                + R.c('importe') + '*' + R.c('tipo') + '/(1-(1+'
                                + R.c('tipo') + ')^-(' + R.c('plazo') + '-'
                                + R.c('carencia') + '))))'),
                          'C': '="Sistema francés, anualidad algebraica. Con '
                               'el tipo al 0 % es el principal entre los años '
                               'de amortización: sin esa rama el cuadro se '
                               'apagaba entero"'})
        rej.add(rot='CUADRO DE AMORTIZACIÓN', bold=True)
        rej.add('cab_cuadro', rot='Año', bold=True,
                valores={'B': 'Capital pendiente', 'C': 'Intereses',
                         'D': 'Amortización de principal', 'E': 'Cuota total',
                         'F': 'Capital al cierre',
                         # A2 / R-14 — la etiqueta dice las dos cosas que hay
                         # que saber para leer la columna: que va después de
                         # impuestos y que a partir del año 4 repite el flujo
                         # del año 3 (el P&L sólo proyecta tres).
                         'G': 'Flujo disponible para la deuda (después de '
                              'impuestos; del año 4 en adelante, el del año 3 '
                              'mantenido)',
                         'H': 'DSCR del año'})
        # RT-07 — el cuadro tenía 10 filas fijas y el plazo no tenía tope de
        # validación: con un plazo de 12 años —el estándar de una línea ICO
        # para inversión en hostelería— quedaban 25.570 € sin amortizar y la
        # hoja acusaba al usuario de haberse equivocado. Ahora llega hasta
        # `motor.PLAZO_MAX` y la validación de datos acota el plazo a ese
        # mismo número con un mensaje que lo dice.
        plazo_max = motor.PLAZO_MAX
        for i in range(1, plazo_max + 1):
            clave = 'y_%d' % i
            anterior = 'y_%d' % (i - 1)
            # el flujo de caja disponible de los años 1-3 sale del P&L; a
            # partir del 4 se mantiene el del año 3, y la nota lo dice: NO se
            # inventa un crecimiento que el modelo no proyecta
            col_pyg = ('B', 'C', 'D')[min(i, 3) - 1]
            rej.add(clave, valores={'A': i}, fmt=motor.FMT_EUR,
                    fmt_A=motor.FMT_ENT, fmt_H=motor.FMT_DEC2,
                    formulas={
                        # pasada la última anualidad el cuadro se APAGA
                        # entero (decisión 14): sin el guarda del plazo, la
                        # fila del año 8 imprimía «0 €» de capital pendiente
                        # en un préstamo a 7 años
                        'B': ((lambda R: '=' + R.c('importe')) if i == 1 else
                              (lambda R, a=anterior, n=i: '=IF(OR(' + str(n)
                               + '>' + R.c('plazo') + ',' + R.c(a, 'F')
                               + '=""),"",' + R.c(a, 'F') + ')')),
                        'C': (lambda R, k=clave, n=i: '=IF(OR(' + str(n) + '>'
                              + R.c('plazo') + ',' + R.c(k) + '=""),"",'
                              + R.c(k) + '*' + R.c('tipo') + ')'),
                        'D': (lambda R, k=clave, n=i: '=IF(OR(' + str(n) + '>'
                              + R.c('plazo') + ',' + R.c(k) + '=""),"",IF('
                              + str(n) + '<=' + R.c('carencia') + ',0,MIN('
                              + R.c(k) + ',' + R.c('cuota') + '-' + R.c(k, 'C')
                              + ')))'),
                        'E': (lambda R, k=clave: '=IF(' + R.c(k, 'C')
                              + '="","",' + R.c(k, 'C') + '+' + R.c(k, 'D')
                              + ')'),
                        'F': (lambda R, k=clave: '=IF(' + R.c(k, 'D')
                              + '="","",' + R.c(k) + '-' + R.c(k, 'D') + ')'),
                        # A2 / R-20 — el flujo disponible para la deuda va
                        # DESPUÉS de impuestos: el numerador era el resultado
                        # ANTES de impuestos con la amortización y los
                        # intereses devueltos, sin descontar el Impuesto de
                        # Sociedades que el propio P&L liquida. El CFADS que
                        # mira una entidad se calcula después del impuesto.
                        'G': (lambda R, k=clave, c=col_pyg: '=IF('
                              + R.c(k, 'E') + '="","",' + pyg.r('rai', c) + '-'
                              + pyg.r('is', c) + '+'
                              + pyg.r('cf_amort', c) + '+' + pyg.r('cf_int', c)
                              + ')'),
                        'H': (lambda R, k=clave: '=IF(OR(' + R.c(k, 'E')
                              + '="",' + R.c(k, 'E') + '=0),"",' + R.c(k, 'G')
                              + '/' + R.c(k, 'E') + ')')})
        rej.add('tolerancia', rot='Tolerancia de cierre del cuadro (€)',
                fmt=motor.FMT_EUR, verdes=('B',),
                valores={'B': 0.01,
                         'C': 'Margen por redondeo. Por encima, el cuadro no '
                              'cierra y hay capital sin devolver'})
        # RC-08 / RT-06 — el control de cierre era INERTE: leía siempre la
        # fila del año 10, que para cualquier plazo menor vale «» y devolvía
        # 0 por construcción, no porque el cuadro cerrase. Ahora se compara
        # contra el PRINCIPAL: lo devuelto tiene que ser lo prestado.
        rej.add('cierre', rot='Capital pendiente al vencimiento',
                fmt=motor.FMT_EUR,
                formulas={'B': (lambda R: '=IF(' + R.c('importe')
                                + '="","",ROUND(' + R.c('importe') + '-SUM('
                                + R.c('y_1', 'D') + ':'
                                + R.c('y_%d' % plazo_max, 'D') + '),2))'),
                          'C': '="Tiene que ser cero: el principal devuelto en '
                               'el cuadro tiene que ser el prestado. Si sale '
                               'positivo, el plazo no llega"'})
        motor.semaforo_num(ws, rej.c('cierre') + ':' + rej.c('cierre'),
                           verde_si='ABS(' + rej.c('cierre') + ')<='
                           + rej.c('tolerancia', absoluta=True),
                           rojo_si='ABS(' + rej.c('cierre') + ')>'
                           + rej.c('tolerancia', absoluta=True))
        # las tres celdas que lee el P&L y la tesorería
        for i in (1, 2, 3):
            rej.add('int_%d' % i, rot='Intereses del año ' + str(i),
                    fmt=motor.FMT_EUR0,
                    formulas={'B': (lambda R, n=i: '=IF(' + R.c('importe')
                                    + '="","",IF(' + R.c('y_%d' % n, 'C')
                                    + '="",0,' + R.c('y_%d' % n, 'C') + '))')})
            rej.add('cap_%d' % i, rot='Devolución de principal del año '
                    + str(i), fmt=motor.FMT_EUR0,
                    formulas={'B': (lambda R, n=i: '=IF(' + R.c('importe')
                                    + '="","",IF(' + R.c('y_%d' % n, 'D')
                                    + '="",0,' + R.c('y_%d' % n, 'D') + '))')})
        rej.add(rot='COBERTURA DEL SERVICIO DE LA DEUDA (DSCR)', bold=True)
        # RD-21 — el DSCR del año 1 es un artefacto de la CARENCIA (sólo hay
        # intereses en el denominador), no una medida del negocio: la cifra
        # que un banco mira es la del primer año de amortización. Se rotula
        # como lo que es y se publica el mínimo de TODO el cuadro, que llega
        # al vencimiento y no se queda en el año 3.
        for i, col in enumerate(('B', 'C', 'D'), start=1):
            rej.add('dscr_%d' % i, rot='DSCR del año ' + str(i),
                    fmt=motor.FMT_DEC2,
                    formulas={'B': (lambda R, n=i, c=col:
                                    '=(' + pyg.r('rai', c) + '-'
                                    + pyg.r('is', c) + '+'
                                    + pyg.r('cf_amort', c) + '+'
                                    + R.c('int_%d' % n) + ')/('
                                    + R.c('int_%d' % n) + '+'
                                    + R.c('cap_%d' % n) + ')'),
                              'C': (('="Año de CARENCIA si la hay: con sólo '
                                     'intereses en el denominador el ratio se '
                                     'dispara y no mide el negocio"')
                                    if i == 1 else None)})
        rej.add('dscr_min', rot='DSCR mínimo de todo el cuadro',
                fmt=motor.FMT_DEC2, bold=True,
                formulas={'B': (lambda R: '=IF(COUNT(' + R.c('y_1', 'H') + ':'
                                + R.c('y_%d' % plazo_max, 'H') + ')=0,"",MIN('
                                + R.c('y_1', 'H') + ':'
                                + R.c('y_%d' % plazo_max, 'H') + '))'),
                          'C': '="El peor año de los que dura el préstamo, no '
                               'el mejor de los tres que proyecta el P&L. Es '
                               'el número que decide una operación. Ojo al '
                               'alcance: los años 1 a 3 salen del P&L y del '
                               'año 4 en adelante el cuadro MANTIENE el flujo '
                               'del año 3, así que el mínimo de esos años es '
                               'una proyección sostenida, no un año '
                               'calculado. El flujo va después de impuestos"'})
        dscr_min = P('dscr_min')
        dscr_obj = P('dscr_obj')
        for clave in ['dscr_%d' % i for i in (1, 2, 3)] + ['dscr_min']:
            # RD-22 — los umbrales del DSCR vivían DENTRO de la fórmula del
            # formato condicional, contra §7-bis.11: el comprador no podía
            # ajustar el ratio al covenant de su banco sin editar reglas de
            # formato. Ahora salen de dos celdas de la hoja de Supuestos.
            motor.semaforo_num(ws, rej.c(clave) + ':' + rej.c(clave),
                               verde_si=rej.c(clave) + '>=' + dscr_obj,
                               ambar_si=rej.c(clave) + '>=' + dscr_min,
                               rojo_si=rej.c(clave) + '<' + dscr_min)
        self.pendientes.append(rej)
        fila = rej.ultima + 2
        motor.val(ws, 'A' + str(fila),
                  'El DSCR se calcula ANTES de la deuda y DESPUÉS de '
                  'impuestos: al resultado antes de impuestos se le resta el '
                  'Impuesto de Sociedades y se le devuelven la amortización '
                  'contable y los intereses; el resultado se divide entre lo '
                  'que hay que pagar ese año. Por debajo del mínimo de la hoja de '
                  'Supuestos, el negocio no genera para pagar el préstamo. '
                  'En el cuadro, el flujo de caja disponible de los años 4 en '
                  'adelante se mantiene en el del año 3: el P&L sólo proyecta '
                  'tres, y aquí no se inventa un crecimiento que no está '
                  'calculado.', wrap=True)
        motor.anchos(ws, {'A': 46, 'B': 18, 'C': 52, 'D': 22, 'E': 16,
                          'F': 18, 'G': 20, 'H': 14})
        motor.print_setup(ws, header_row=cab)
        return rej


    # -- §2.9 -------------------------------------------------------------
    # -- §2.9 -------------------------------------------------------------
    def instrucciones(self):
        """`Instrucciones`: textos ciertos y ratios que auditan (§2.9)."""
        ws = self.ws_ins
        col = motor._col_texto(ws)
        letra = get_column_letter(col)
        # se limpia sólo el cuerpo: `motor.cierre_instrucciones()` reescribe
        # después su bloque (desproteger + cross-sell + bio + versión)
        _limpiar_area(ws, 1, ws.max_row, col + 4)
        motor.val(ws, letra + '1', 'INSTRUCCIONES DE USO — Plan financiero '
                  + str(self.concepto), bold=True)
        fila = 3
        pyg = self.rej['pyg']
        inv = self.rej['inversion']
        eq = self.rej['equilibrio']
        lineas = [
            'CÓMO SE USA ESTE LIBRO',
            'Este plan financiero es un MODELO: sólo se teclea en las celdas '
            'VERDES y el resto se recalcula solo.',
            '1. «0. Supuestos» es la hoja de mando: cubiertos, ticket sin '
            'IVA, días de apertura, mezcla de comida y bebida, coste de '
            'mercancía, alquiler, financiación e impuestos.',
            '2. Las partidas de la inversión, los costes fijos y la plantilla '
            'se teclean en su propia hoja, también en verde. Ningún número se '
            'escribe dos veces.',
            # RC-05 — decía que cuatro hojas «no hay que tocarlas» cuando
            # entre las tres primeras suman 26 celdas VERDES con validación
            # de datos, que por la convención del propio fichero son
            # justamente lo que hay que teclear.
            # REF-10 (`…tapas-bar-ref.json`, 2026-08-29): este párrafo citaba
            # los nombres de hoja NUMERADOS del molde A-α aunque el libro
            # fuera A-β (hojas SIN numerar): se compone con el título REAL
            # de cada hoja, ya resuelto en `__init__` (`self.ws_<clave>`),
            # para que valga en los dos moldes.
            # ⚠️ M8 / R22-TAP-11 / R22-CAF-06 (2026-09-05). El punto 3
            # afirmaba que la hoja de Punto de Equilibrio «se deriva ENTERA»
            # cuando el motor 2.2 le puso una celda VERDE —el techo de
            # rotaciones del local— de la que cuelga un semáforo: quien
            # siguiera las instrucciones no la tocaría y compararía su plan
            # contra un 3 que no es el suyo. La frase se compone según el
            # molde, porque los moldes de mostrador (panadería, food truck)
            # apagan el bloque y ahí la afirmación original SÍ es cierta.
            ('3. En «' + self.ws_equilibrio.title + '» sólo se teclea el '
             'techo de rotaciones que da tu local; todo lo demás se deriva '
             'de lo anterior.'
             if (self.dato('ROTACION', {}) or {}).get('activa', True)
             else '3. Sólo «' + self.ws_equilibrio.title + '» se deriva '
                  'entera de lo anterior.')
            + ' En las otras tres hojas también hay celdas verdes que '
            'puedes tocar: el escenario pesimista y el optimista en «'
            + self.ws_escenarios.title + '», la estacionalidad y la rampa '
            'de arranque en «' + self.ws_tesoreria.title + '», y las '
            'cuatro fuentes alternativas de financiación en «'
            + self.ws_financiacion.title + '». Todo lo demás se calcula.',
            # ⚠️ M19 / REF22-BAR-03 / FT22-09 (2026-09-05). Dos arreglos:
            # (a) el divisor 1,04 se nombraba en los CINCO hermanos y sólo la
            # panadería tiene una línea al tipo superreducido — en el bar es
            # ruido que describe una línea de venta que no existe; ahora se
            # menciona sólo si alguna línea del libro lo declara. (b) la
            # bollería y la pastelería VENDIDAS PARA LLEVAR son una entrega
            # de bienes y su 10 % sale del art. 91.Uno.1.1.º, no del
            # 91.Uno.2.2.º, que es el de los SERVICIOS de hostelería: se
            # citaba el artículo equivocado justo en el único supuesto en el
            # que la mención aportaba algo.
            '4. Todas las cifras van SIN IVA. Para pasar un PVP a precio sin '
            'IVA, divide por el tipo que le corresponda: '
            + ('entre 1,04 el pan común y las harinas panificables (tipo '
               'superreducido, art. 91.Dos.1.1.º de la Ley 37/1992 del IVA); '
               'entre 1,10 todo el consumo en el acto —comida y bebida, '
               'alcohol incluido— (art. 91.Uno.2.2.º) y, cuando se venden '
               'para llevar, los demás alimentos, bollería y pastelería '
               'incluidas, que son productos alimenticios al tipo reducido '
               'por el art. 91.Uno.1.1.º; '
               if any(abs(t - 0.04) < 1e-9 for t in
                      (getattr(self, 'tipos_especiales', None) or ()))
               else 'entre 1,10 en sala (comida y bebida, alcohol incluido: '
                    'art. 91.Uno.2.2.º de la Ley 37/1992 del IVA) y, cuando '
                    'se venden para llevar, los alimentos —bollería y '
                    'pastelería incluidas—, que son productos alimenticios '
                    'al tipo reducido por el art. 91.Uno.1.1.º; ')
            + 'y entre 1,21 el tipo general, que en hostelería sólo alcanza '
            'a lo que sale del local como entrega de bienes excluida del '
            'reducido: el alcohol y los refrescos o zumos con azúcares '
            'añadidos para llevar o a domicilio (la comida para llevar sigue '
            'al 10 %); el libro lo aplica al alcohol del canal de delivery. '
            'El tipo de cada línea de venta está en la columna «Tipo de IVA '
            'de la línea» del P&L y de ahí salen los cobros, el IVA '
            'repercutido y el PVP equivalente. El IVA de la inversión se '
            'adelanta y se recupera con el modelo 303.',
            '5. Si cambias la plantilla, el P&L la lee sola: el coste de '
            'personal sale de la hoja «Personal», no de una estimación '
            'aparte.',
        ]
        for texto in (self.dato('INSTRUCCIONES', {}) or {}).get('uso', []):
            lineas.append(texto)
        for texto in lineas:
            motor.val(ws, letra + str(fila), texto, wrap=True,
                      bold=texto.isupper())
            fila += 1
        fila += 1
        motor.val(ws, letra + str(fila), 'RATIOS QUE AUDITA EL LIBRO',
                  bold=True)
        fila += 1
        motor.val(ws, letra + str(fila),
                  'Los umbrales viven en la columna «Umbral» del bloque '
                  'RATIOS CLAVE del P&L: cámbialos ahí y el semáforo cambia '
                  'con ellos.', wrap=True)
        fila += 2
        cabecera = fila
        for i, texto in enumerate(('Ratio', 'Valor del año 1', 'Umbral',
                                   'Estado')):
            motor.val(ws, get_column_letter(col + i) + str(fila), texto,
                      bold=True)
        fila += 1
        # A6 — el mismo rótulo que publica el bloque RATIOS CLAVE del P&L:
        # si el molde lo renombra, las dos tablas tienen que decir lo mismo.
        _umbrales_rot = dict((u[0], u[1])
                             for u in (self.dato('UMBRALES') or ()) if u[1])
        for clave, rot in (('r_cogs', 'Coste de mercancía / Ventas'),
                           ('r_personal', 'Coste de personal / Ventas'),
                           ('r_alquiler', 'Alquiler / Ventas'),
                           ('r_neto', 'Resultado neto / Ventas'),
                           ('r_mb', 'Margen bruto / Ventas')):
            motor.val(ws, letra + str(fila), _umbrales_rot.get(clave, rot))
            fx(ws, get_column_letter(col + 1) + str(fila),
               '=' + pyg.r(clave), motor.FMT_PCT)
            fx(ws, get_column_letter(col + 2) + str(fila),
               '=' + pyg.r(clave, 'E'), motor.FMT_PCT)
            menor = clave in ('r_cogs', 'r_personal', 'r_alquiler')
            c1 = get_column_letter(col + 1) + str(fila)
            c2 = get_column_letter(col + 2) + str(fila)
            fx(ws, get_column_letter(col + 3) + str(fila),
               '=IF(OR(' + c1 + '="",' + c2 + '=""),"",IF(' + c1
               + ('<=' if menor else '>=') + c2 + ',"CUMPLE","REVISAR"))')
            fila += 1
        motor.semaforo_texto(
            ws, get_column_letter(col + 3) + str(cabecera + 1) + ':'
            + get_column_letter(col + 3) + str(fila - 1),
            (('CUMPLE', motor.CF_VERDE_BG, motor.CF_VERDE_FG),
             ('REVISAR', motor.CF_ROJO_BG, motor.CF_ROJO_FG)))
        fila += 1
        motor.val(ws, letra + str(fila), 'CIFRAS DE ESTE PLAN', bold=True)
        fila += 1
        for rot, formula, fmt in (
                ('Inversión total (suma de los bloques)',
                 '=' + inv.r('total'), motor.FMT_EUR0),
                ('Necesidad total de caja al arranque',
                 '=' + inv.r('caja'), motor.FMT_EUR0),
                ('Fondo de maniobra dotado', '=' + inv.r('fondo'),
                 motor.FMT_EUR0),
                ('Facturación prevista del año 1', '=' + pyg.r('ingresos'),
                 motor.FMT_EUR0),
                ('Resultado neto del año 1', '=' + pyg.r('neto'),
                 motor.FMT_EUR0),
                # M1 — la fila publicaba «el» punto de equilibrio y hay DOS:
                # el rótulo dice cuál es (el contable, `cub_dia`).
                ('Cubiertos/día para el punto de equilibrio CONTABLE (con la '
                 'amortización dentro)', '=' + eq.r('cub_dia'),
                 motor.FMT_DEC),
                ('Payback del proyecto (años), antes de la deuda',
                 '=' + self.rej['tesoreria'].r('payback'), 'General')):
            motor.val(ws, letra + str(fila), rot)
            fx(ws, get_column_letter(col + 1) + str(fila), formula, fmt)
            fila += 1
        # RD-07 — el caso base pasa de perder 22.170 € a ganar 40.131 €
        # accionando a la vez dos palancas (+45 % de cubiertos y −29,7 % de
        # nómina) y el fichero no dejaba rastro de qué se había movido ni por
        # qué. Aquí queda, con la cifra vieja, la nueva y su justificación.
        recalibrado = self.dato('RECALIBRADO', []) or []
        if recalibrado:
            fila += 1
            motor.val(ws, letra + str(fila),
                      'QUÉ HA CAMBIADO RESPECTO DE LA VERSIÓN 1.1',
                      bold=True)
            fila += 1
            for i, texto in enumerate(('Concepto', 'v1.1',
                                       'v' + motor.VERSION, 'Por qué')):
                motor.val(ws, get_column_letter(col + i) + str(fila), texto,
                          bold=True)
            fila += 1
            for entrada in recalibrado:
                partes = (list(entrada) + [''] * 4)[:4]
                for i, texto in enumerate(partes):
                    # openpyxl toma por FÓRMULA cualquier cadena que empiece
                    # por «=», y `inject_cache` la marca luego como fórmula
                    # sin caché: un texto explicativo no puede abrir con «=»
                    if isinstance(texto, str) and texto.startswith('='):
                        texto = texto.lstrip('=').strip()
                    motor.val(ws, get_column_letter(col + i) + str(fila),
                              texto, wrap=(i == 3))
                fila += 1
        fila += 1
        motor.val(ws, letra + str(fila),
                  'DATOS DE REFERENCIA DEL SECTOR', bold=True)
        fila += 1
        for i, texto in enumerate(('Referencia', 'Valor', 'Fuente',
                                   'Nota')):
            motor.val(ws, get_column_letter(col + i) + str(fila), texto,
                      bold=True)
        fila += 1
        # RD-33 — cuatro de las cinco filas llevaban su procedencia y la del
        # ticket medio no, pese a venir del mismo sitio: en el dato que más
        # se discute con un banco, el que no decía de dónde salía era
        # precisamente el ticket. La fuente pasa a ser una columna propia.
        for entrada in (self.dato('INSTRUCCIONES', {}) or {}).get(
                'referencias', []):
            rot, valor, fuente, nota = (list(entrada) + [''] * 4)[:4]
            motor.val(ws, letra + str(fila), rot)
            motor.val(ws, get_column_letter(col + 1) + str(fila), valor)
            motor.val(ws, get_column_letter(col + 2) + str(fila), fuente)
            motor.val(ws, get_column_letter(col + 3) + str(fila), nota,
                      wrap=True)
            fila += 1
        motor.anchos(ws, {letra: 62,
                          get_column_letter(col + 1): 18,
                          get_column_letter(col + 2): 22,
                          get_column_letter(col + 3): 54})
        motor.print_setup(ws)
        return fila

    # -- §2.10 ------------------------------------------------------------
    def checklist(self, ws_libro):
        """Checklist de apertura: legal vigente y sin inventos (§2.10)."""
        reglas = self.dato('CHECKLIST', {}) or {}
        reemplazos = reglas.get('reemplazos', [])
        # M5 — los patrones que NO encuentran celda en crudo se reintentan
        # contra el texto con el saneado del §1 deshecho. El censo se hace
        # una sola vez, sobre el libro entero, ANTES de tocar nada: así un
        # reemplazo que ya funciona no puede cambiar de destino.
        _textos0 = _textos_de(ws_libro)
        _crudos = _casan_en_crudo(_textos0, reemplazos)
        _claves = [motor.norm(pat) for pat, _n in reemplazos]
        altas = reglas.get('altas', [])
        suprimir = [motor.norm(s) for s in reglas.get('suprimir', [])]
        fases = reglas.get('fases', {})
        tocados, anadidos = 0, 0
        for ws in ws_libro.worksheets:
            if motor.norm(ws.title) in motor.HOJAS_MOTOR:
                continue
            cab = _cabecera(ws)
            cols = dict((motor.norm(c.value), c.column) for c in ws[cab]
                        if isinstance(c.value, str))
            # T7/panadería (2026-08-29): el checklist de este hermano
            # rotula la columna simplemente «TRAMITE» —sin «/ Tarea» ni
            # «/ Accion»—, un cuarto texto de cabecera que ninguno de los
            # tres anteriores usaba. Sin `cols.get('tramite')` el método
            # entero se saltaba la hoja (`continue`) — reemplazos Y altas,
            # las dos cosas, en las 6 pestañas — SIN un solo aviso: el
            # dry-run daba TODO VERDE porque ningún gate mide «¿se
            # ejecutaron los reemplazos?», sólo que las fórmulas nuevas
            # evalúen. Se añade como ÚLTIMO recurso (después de las tres
            # variantes ya vistas) para no cambiar cuál gana si un fichero
            # futuro trajera dos cabeceras candidatas.
            col_tarea = (cols.get('tramite / tarea') or cols.get('tarea')
                         or cols.get('tramite / accion') or cols.get('hito')
                         or cols.get('tramite'))
            if col_tarea is None:
                continue
            # REF-04 (`…tapas-bar-ref.json`, 2026-08-29): en el molde C2 la
            # columna «OK» (la casilla ☐/✓/N/A) va ANTES de la de tarea, y
            # las altas se escribían sin nada en ella — el desplegable que
            # `motor.checklist_ok_y_contador` añade después las deja
            # marcables, pero en blanco, distintas a simple vista de las
            # filas heredadas.
            col_ok = cols.get('ok')
            # (a) reemplazos de contenido, celda a celda
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cel = ws.cell(row=r, column=c)
                    if not isinstance(cel.value, str):
                        continue
                    # M5 — el texto con el saneado del §1 deshecho
                    _kv = clave_reemplazo(cel.value)
                    for _i, (patron, nuevo) in enumerate(reemplazos):
                        if re.search(patron, cel.value, re.I) \
                                or (_i not in _crudos
                                    and re.search(_claves[_i], _kv, re.I)):
                            if cel.value != nuevo:
                                self.anota(ws.title + '!' + cel.coordinate
                                           + ': «' + cel.value[:48]
                                           + '» → «' + nuevo[:48] + '»')
                                cel.value = nuevo
                                tocados += 1
                            break
            # (b) cabeceras de fase con el cronograma ÚNICO (DOM-23)
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if not isinstance(v, str):
                    continue
                for patron, nuevo in fases.items():
                    if re.search(patron, v, re.I) and v != nuevo:
                        self.anota(ws.title + '!A' + str(r) + ': cronograma «'
                                   + v[:44] + '» → «' + nuevo[:44] + '»')
                        ws.cell(row=r, column=1).value = nuevo
                        tocados += 1
            # (c) filas suprimidas (duplicados y trámites que no aplican)
            if suprimir:
                # RD-30 / RC-22 — vaciar la fila dejaba un hueco con bordes y
                # alto de 22 pt en mitad de la FASE 2, entre el Registro
                # Sanitario y la licencia de terraza: se lee como una tarea
                # que alguien se ha olvidado de rellenar en el fichero que se
                # vende como «no te dejas nada pendiente». Se BORRA.
                for r in range(ws.max_row, cab, -1):
                    v = ws.cell(row=r, column=col_tarea).value
                    if isinstance(v, str) and motor.norm(v) in suprimir:
                        _borrar_fila(ws, r)
                        tocados += 1
                        self.anota(ws.title + ': fila ' + str(r)
                                   + ' BORRADA (no vaciada) — ' + v[:60])
            # (d) altas al final de su fase
            destino = altas if len(ws_libro.worksheets) <= 2 else [
                a for a in altas if re.search(a[0], ws.title, re.I)]
            if destino:
                anadidos += self._altas_checklist(
                    ws, cab, col_tarea, destino,
                    cabecera=reglas.get('cabecera_altas'),
                    rx_fase_final=reglas.get('fase_final'),
                    fase_final_nueva=reglas.get('fase_final_nueva'),
                    col_ok=col_ok)
        # M5 / R22-TAP-07 — el gate que faltaba desde REF-09
        no_entregados, muertos = auditar_reemplazos(ws_libro, reemplazos)
        for patron, nuevo, celda in no_entregados:
            self.anota('⚠️ FALLO Checklist: el reemplazo ' + repr(patron)
                       + ' alcanza «' + str(celda)[:50] + '» y NO se aplicó: '
                       'el libro se entrega SIN «' + nuevo[:60]
                       + '» (M5 / R22-TAP-07)')
        for patron, nuevo, _c in muertos:
            self.anota('⚠️ Checklist: el reemplazo ' + repr(patron) + ' no '
                       'llega a ninguna celda de este fichero (fila que no '
                       'existe o que otro reemplazo anterior ya renombró): '
                       'no se entrega «' + nuevo[:60] + '» (M5, aviso)')
        self.anota('Checklist: ' + str(tocados) + ' celdas corregidas, '
                   + str(anadidos) + ' trámites nuevos, '
                   + str(len(no_entregados)) + ' reemplazos no entregados y '
                   + str(len(muertos)) + ' patrones muertos (§2.10)')
        return tocados, anadidos

    def _altas_checklist(self, ws, cab, col_tarea, altas, cabecera=None,
                         rx_fase_final=None, fase_final_nueva=None,
                         col_ok=None):
        """Añade trámites conservando la estructura de fases del fichero.

        IDEMPOTENTE por CONTENIDO: los que ya están no se vuelven a añadir. Sin
        esto la 2.ª pasada duplicaba los diez trámites nuevos y el contador del
        checklist pasaba de 59 a 69 sin que nadie lo hubiera pedido.
        """
        # REF-04 — la marca de «sin marcar» se COPIA de una fila heredada
        # (nunca se inventa `motor.VOCAB_MARCA[1]`, que es «—»: los ficheros
        # de esta familia ya traen su propio vocabulario, «☐»/«✓»/«N/A», y
        # escribir un símbolo distinto del que usa el resto de la hoja
        # rompería la lectura visual de las filas heredadas).
        marca_vacia = None
        if col_ok:
            for r in range(cab + 1, ws.max_row + 1):
                v = ws.cell(row=r, column=col_ok).value
                if isinstance(v, str) and v.strip():
                    marca_vacia = v
                    break
        existentes = set()
        for r in range(cab + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_tarea).value
            if isinstance(v, str) and v.strip():
                existentes.add(motor.norm(v))
        altas = [a for a in altas if motor.norm(a[2]) not in existentes]
        if not altas:
            return 0
        # última fila con contenido en la columna de tarea, SIN contar el pie
        # «Tareas completadas: X de Y» (REF-04/REF-05,
        # `auditorias/planes-v2-hermano-plan-negocio-tapas-bar-ref.json`,
        # 2026-08-29): si el pie queda dentro de `ultima`, las altas se
        # insertan DEBAJO de él —sin su propia casilla— y el rango
        # `letra_ok+r0:letra_ok+r1` de `motor.checklist_ok_y_contador` (que sí
        # excluye el pie de `filas`, pero no del RANGO r0:r1) pasa a contar la
        # etiqueta del pie como si fuera una tarea más. Misma exclusión
        # ('completad' en el rótulo normalizado) que usa el motor, para que
        # los dos lean el mismo pie.
        ultima = cab
        for r in range(cab + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_tarea).value
            if isinstance(v, str) and v.strip() \
                    and 'completad' not in motor.norm(v):
                ultima = r
        # RD-28 / RC-12 — la fase que hay que dejar la ÚLTIMA (los primeros
        # 90 días de operación) se captura, se vacía y se reescribe DESPUÉS
        # del bloque nuevo: si no, las obligaciones que hay que tener
        # cerradas antes de abrir quedan detrás de las de la operación ya en
        # marcha, y quien trabaja el checklist de arriba abajo contrata el
        # seguro de RC con el local abierto.
        cola, fila_fase = [], None
        if rx_fase_final:
            for r in range(cab + 1, ultima + 1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str) and re.search(rx_fase_final, v, re.I):
                    fila_fase = r
                    break
        if fila_fase is not None:
            for r in range(fila_fase, ultima + 1):
                fila = [(c, ws.cell(row=r, column=c).value)
                        for c in range(1, ws.max_column + 1)
                        if ws.cell(row=r, column=c).value is not None]
                cola.append((r, fila, _ancho_combinado(ws, r)))
            for r, fila, _a in cola:
                for m in list(ws.merged_cells.ranges):
                    cr = CellRange(str(m))
                    if cr.min_row <= r <= cr.max_row:
                        _unmerge(ws, m)
                for c, _v in fila:
                    ws.cell(row=r, column=c).value = None
            ultima = fila_fase - 1
        pie = []
        for r in range(ultima + 1 + len(cola), ws.max_row + 1):
            fila = [(c, ws.cell(row=r, column=c).value)
                    for c in range(1, ws.max_column + 1)
                    if ws.cell(row=r, column=c).value is not None]
            if fila:
                pie.append((r, fila, _ancho_combinado(ws, r)))
        # el pie va COMBINADO a lo ancho de la tabla: si se mueve sin
        # deshacer la combinación, `MergedCell.value` es de sólo lectura y la
        # escritura revienta con AttributeError
        for r, fila, _ancho in pie:
            for m in list(ws.merged_cells.ranges):
                cr = CellRange(str(m))
                if cr.min_row <= r <= cr.max_row:
                    _unmerge(ws, m)
            for c, _v in fila:
                ws.cell(row=r, column=c).value = None
        cols = dict((motor.norm(c.value), c.column) for c in ws[cab]
                    if isinstance(c.value, str))
        # el cuerpo se reescribe entero por debajo de `ultima`: cualquier
        # combinada que sobreviva ahí convierte la celda destino en
        # `MergedCell`, cuyo `.value` es de SÓLO LECTURA. Se deshacen todas
        # de una vez y las que hacen falta se rehacen al escribir.
        for m in list(ws.merged_cells.ranges):
            if CellRange(str(m)).max_row > ultima:
                _unmerge(ws, m)
        fila = ultima + 1
        n = 0
        usa_fase = 'fase' in cols
        if cabecera and altas:
            _desmerge_fila(ws, fila)
            motor.val(ws, 'A' + str(fila), cabecera, bold=True)
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                           end_column=ws.max_column)
            fila += 1
        for _hoja, fase, tarea, responsable, plazo, nota in altas:
            _desmerge_fila(ws, fila)
            valores = {col_tarea: tarea}
            if col_ok and marca_vacia:
                valores[col_ok] = marca_vacia
            if usa_fase:
                valores[cols['fase']] = fase
            if 'responsable' in cols:
                valores[cols['responsable']] = responsable
            if 'plazo' in cols:
                valores[cols['plazo']] = plazo
            if 'notas' in cols:
                valores[cols['notas']] = nota
            for c, v in valores.items():
                motor.val(ws, get_column_letter(c) + str(fila), v, wrap=True)
            fila += 1
            n += 1
        # la fase que va la última se reescribe aquí, ya por debajo
        for _r, contenido, ancho in cola:
            _desmerge_fila(ws, fila)
            for c, v in contenido:
                if c == 1 and fase_final_nueva and isinstance(v, str) \
                        and rx_fase_final and re.search(rx_fase_final, v,
                                                        re.I):
                    v = fase_final_nueva
                motor.val(ws, get_column_letter(c) + str(fila), v, wrap=True)
            if ancho > 1:
                ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                               end_column=ancho)
                ws[get_column_letter(1) + str(fila)].font = Font(bold=True)
            fila += 1
        fila += 1
        for _r, contenido, ancho in pie:
            _desmerge_fila(ws, fila)
            for c, v in contenido:
                motor.val(ws, get_column_letter(c) + str(fila), v)
            if ancho > 1:
                ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                               end_column=ancho)
            fila += 1
        # RT-01 — el resaltado de fila del molde C1 se reancla al cuerpo
        # nuevo. Ojo: hay que reescribir la FÓRMULA, no sólo el `sqref`.
        # Reutilizar el objeto regla dejaba `sqref=A3:F73` con `$E4`, así que
        # la fila 5 (primer ítem) se pintaba según E6 y la CABECERA según E4:
        # el resaltado entero quedaba desplazado una fila, y en el fichero de
        # producción (`sqref=A4:F59` con `$E4`) estaba BIEN. Es una regresión
        # del motor, y por eso `motor.gate_cf_anclado` la vigila.
        for cf in list(ws.conditional_formatting):
            reglas = list(cf.rules)
            if not reglas or reglas[0].type != 'expression':
                continue
            regla = reglas[0]
            formula = str((regla.formula or [''])[0])
            primera = cab + 1
            reanclada = re.sub(r'(\$?[A-Z]{1,3}\$?)(\d+)',
                               lambda m: m.group(1) + str(primera), formula)
            nueva = Rule(type='expression', formula=[reanclada],
                         dxf=copy.copy(regla.dxf), stopIfTrue=regla.stopIfTrue)
            # ⚠️ M18 / R22-TAP-14 / FT22-21 (2026-09-05). El rango
            # terminaba en `fila`, que después de reescribir la cola de
            # fases y el pie queda DOS o TRES filas por debajo del último
            # trámite: cubría la fila de totales y una vacía. Hoy no pinta
            # nada (ninguna tiene ✓ en la columna A), pero un ítem añadido
            # ahí nacería con formato de «completado». Se recalcula desde la
            # última fila de ítems REAL, la misma que usa el contador.
            ultimo_item = primera - 1
            for r in range(primera, ws.max_row + 1):
                v = ws.cell(row=r, column=col_tarea).value
                if isinstance(v, str) and v.strip() \
                        and 'completad' not in motor.norm(v) \
                        and not motor.RX_TOTAL.match(str(v)):
                    ultimo_item = r
            ultimo_item = max(ultimo_item, primera)
            _purgar_cf_area(ws, cab, ws.max_row)
            ws.conditional_formatting.add(
                'A' + str(primera) + ':'
                + get_column_letter(ws.max_column) + str(ultimo_item), nueva)
            self.anota(ws.title + ': resaltado de fila reanclado a A'
                       + str(primera) + ':' + get_column_letter(ws.max_column)
                       + str(ultimo_item) + ' con la fórmula «' + reanclada
                       + '» (RT-01 · M18)')
            break
        return n


# ==========================================================================
# M5 / R22-TAP-07 — el §1 transversal mueve la portería del checklist
# ==========================================================================
#: ⚠️ (2026-09-05) Los `CHECKLIST['reemplazos']` de cada módulo de contenido
#: están escritos contra el texto del fichero de PARTIDA, y `motor.aplicar()`
#: reescribe ese texto ANTES de que corra `grupo_a.checklist()`: tildes
#: (`bano maria` → `baño maría`), erratas (`Priorizarcexperiencia` →
#: `Priorizar experiencia`), `EUR` → `€` (RX_EUR_TEXTO) y `capital social
#: min 1` → `capital social mín. 1` (RX_CAPITAL_MIN). Medido en la refutación
#: de la 2.2: CUATRO reemplazos muertos en tapas-bar y DOS en el
#: bar-restaurante, todos SILENCIOSOS y con el dry-run en 13/13 verde.
#:
#: Arreglar los patrones uno a uno no basta —la próxima palabra que entre en
#: `TILDES` mataría otro reemplazo sin avisar—, así que el motor DESHACE las
#: transformaciones del §1 para comparar. La clave se usa sólo como SEGUNDA
#: oportunidad y sólo para los patrones que no casan en crudo en ninguna
#: celda: un reemplazo que ya funciona nunca cambia de destino por esto.
_ERRATAS_INV = sorted(
    [(motor.norm(bueno), motor.norm(malo)) for malo, bueno in motor.ERRATAS],
    key=lambda kv: -len(kv[0]))


def clave_reemplazo(texto):
    """El texto de una celda, deshecho el saneado del §1 transversal (M5)."""
    t = motor.norm(texto)
    t = t.replace('\u20ac', 'eur')          # § RX_EUR_TEXTO
    t = re.sub(r'\bmin\.', 'min', t)        # § RX_CAPITAL_MIN
    for bueno, malo in _ERRATAS_INV:
        if bueno in t:
            t = t.replace(bueno, malo)
    return t


def _casan_en_crudo(textos, reemplazos):
    """Índices de los patrones que ya encuentran celda sin ayuda."""
    return set(i for i, (pat, _n) in enumerate(reemplazos)
               if any(re.search(pat, v, re.I) for v in textos))


def _textos_de(wb):
    fuera = []
    for ws in wb.worksheets:
        if motor.norm(ws.title) in motor.HOJAS_MOTOR:
            continue
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip():
                    fuera.append(c.value)
    return fuera


def auditar_reemplazos(wb, reemplazos):
    """El gate que faltaba desde REF-09 (M5 / R22-TAP-07).

    Devuelve `(no_entregados, muertos)` sobre el libro YA procesado:

    * **no_entregados** — el texto nuevo NO está en el libro y, sin embargo,
      hay una celda a la que el patrón llega (en crudo o deshaciendo el §1) y
      que ningún reemplazo anterior le había ganado. Es el fallo de verdad:
      el comprador recibe el entregable SIN la corrección. Cuenta como fallo
      del dry-run.
    * **muertos** — el texto nuevo no está y el patrón no llega a ninguna
      celda ni deshaciendo el §1: el módulo declara una corrección para una
      fila que este fichero no tiene (o que otro reemplazo ya renombró). No
      puede estropear nada, así que se AVISA y no se tumba la tanda.

    El reparto por SOMBRA importa: el bar-restaurante declara dos veces
    `^Alta en el IAE` (la segunda es una redacción posterior de la misma
    fila) y sin la exención la segunda parecería un fallo.
    """
    reemplazos = list(reemplazos or [])
    textos = _textos_de(wb)
    presentes = set(textos)
    crudos = _casan_en_crudo(textos, reemplazos)
    claves = [motor.norm(pat) for pat, _n in reemplazos]
    # ganador de cada celda y patrones que también la alcanzaban
    alcanza, gana = {}, {}
    for v in textos:
        kv = clave_reemplazo(v)
        for i, (pat, _n) in enumerate(reemplazos):
            if re.search(pat, v, re.I) or (i not in crudos
                                           and re.search(claves[i], kv,
                                                         re.I)):
                alcanza.setdefault(i, []).append(v)
                if v not in gana:
                    gana[v] = i
    no_entregados, muertos = [], []
    for i, (pat, nue) in enumerate(reemplazos):
        if nue in presentes:
            continue
        celdas = alcanza.get(i) or []
        propias = [v for v in celdas if gana.get(v) == i]
        if propias:
            no_entregados.append((pat, nue, propias[0]))
        else:
            muertos.append((pat, nue, celdas[0] if celdas else None))
    return no_entregados, muertos


# ==========================================================================
# Ganchos que consume `main.py`
# ==========================================================================
def ficheros(dets, contenido=None):
    """Los xlsx de línea A: el plan financiero y el checklist de apertura."""
    fuera = []
    for fname, det in (dets or {}).items():
        if det['tipo'] == 'plan_financiero' and det['molde'] in MOLDES:
            fuera.append(fname)
        elif det['tipo'] == 'checklist' and det['molde'] in ('C1', 'C2'):
            fuera.append(fname)
    return fuera


def _titular(wb, det, plan, cambios):
    """M11 / R22-TAP-16 / FT22-05 — el `dc:title` de docProps.

    Se heredaba el A1/A3 de la hoja de Instrucciones de la v1.1 («RATIOS
    REFERENCIA — TAPAS BAR / GASTROBAR ESPANA 2026 · …»): describe una hoja
    que en la 2.2 ya no se llama así y va SIN eñe, porque el §1.7 acentúa
    celdas y no metadatos. Se compone desde el CONCEPTO del módulo de
    contenido y se pasa por el mismo saneado de tildes. `motor.metadatos`,
    que corre después en `cerrar()`, le añade el «· vX.Y».
    """
    familia = ('Checklist de apertura' if det['tipo'] == 'checklist'
               else 'Plan financiero')
    base = motor.corregir_texto(familia + ' — ' + str(plan.concepto))
    viejo = wb.properties.title
    wb.properties.title = base
    wb.properties.subject = base
    cambios.append('docProps title: «' + str(viejo)[:70] + '» → «' + base
                   + ' · v' + motor.VERSION + '» (M11)')
    return base


def post(wb, fname, det, pid, params, cambios, contenido, carpeta=None):
    """§2 completo, DESPUÉS del §1 transversal y ANTES del cierre del motor."""
    if det['tipo'] == 'checklist':
        plan = Plan(wb, det, pid, params, contenido, cambios)
        _titular(wb, det, plan, cambios)
        plan.checklist(wb)
        return
    if det['molde'] not in MOLDES:
        return
    plan = Plan(wb, det, pid, params, contenido, cambios)
    _titular(wb, det, plan, cambios)
    # ORDEN: los supuestos primero (todo cuelga de ellos) y las hojas nuevas
    # ANTES del P&L, porque el P&L lee los intereses de `Financiación` y el
    # fondo de maniobra sale de los costes fijos del P&L. Las rejillas se
    # construyen en dos fases (declarar y escribir), así que una hoja puede
    # citar coordenadas de otra que todavía no se ha volcado.
    plan.supuestos_altas()
    # ORDEN DE DECLARACIÓN: cada hoja cita coordenadas de las anteriores; las
    # citas «hacia atrás» (el P&L pide los intereses a Financiación, que
    # todavía no existe) van como función y se resuelven en el volcado final.
    plan.personal()
    plan.pyg()
    plan.inversion()
    # `Financiación` se DECLARA antes que el punto de equilibrio: desde
    # RD-05 el break-even de caja lee de ella la devolución de principal del
    # año, y una rejilla sólo se puede citar cuando ya existe (las citas
    # «hacia atrás» son las que van como función).
    plan.financiacion()
    plan.equilibrio()
    plan.escenarios()
    plan.tesoreria()
    for rej in plan.pendientes:
        escribir(rej)
    plan.instrucciones()          # cita celdas de todas las anteriores
    plan.supuestos_calculadas()   # necesita las coordenadas del P&L
    # M9 — la ÚLTIMA pasada: traduce el vocabulario del driver en todo lo que
    # este grupo acaba de escribir. Va después de todo porque los rótulos se
    # componen en varias hojas y algunas se citan entre sí por coordenada,
    # nunca por texto.
    plan.vocabulario()
    recalibrado = plan.dato('RECALIBRADO', []) or []
    for entrada in recalibrado:
        cambios.append('RECALIBRADO · ' + ' · '.join(str(x) for x in entrada))
    cambios.append('§2 aplicado sobre el molde ' + det['molde'] + ' de '
                   + fname)


# ==========================================================================
# Demostraciones §2.11 (bloqueantes)
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref_):
    import contextlib
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref_)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__ + ':' + str(e)[:80]


def _clon(origen, destino, cambios):
    """Copia con los inputs cambiados: nunca se toca el entregable."""
    import openpyxl
    shutil.copy2(origen, destino)
    wb = openpyxl.load_workbook(destino)
    for hoja_, coord, valor in cambios:
        ws = motor.hoja(wb, hoja_, obligatoria=True)
        ws[coord] = valor
    wb.save(destino)
    return destino


def _pf(carpeta):
    for n in sorted(os.listdir(carpeta)):
        if n.startswith('plan-financiero') and n.endswith('.xlsx'):
            return os.path.join(carpeta, n)
    return None


def _mapa(path):
    """Coordenadas por rótulo, para que la demo no dependa de la posición."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    mapa = {}
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            rot = motor._rotulo_de_fila(ws, r, max_col=1)
            if rot:
                mapa.setdefault(motor.norm(ws.title), {}).setdefault(
                    motor.norm(rot), r)
    return wb, mapa


#: Cubiertos/día del ensayo de la demo 5: el año 1 pierde, el año 2 gana
#: menos de lo que arrastra en bases negativas y el año 3 tributa al tipo
#: reducido, así que las tres ramas del impuesto quedan demostradas de una
#: vez. La ventana es estrecha y se mueve con cualquier recalibración del
#: caso base (la abrió y la cerró RD-10 al hacer escalonado el personal), así
#: que el escenario se BUSCA en vez de teclearse: se prueban valores
#: decrecientes y se toma el primero que enseña las tres ramas.
DEMO5_CUBIERTOS = 65
DEMO5_FACTORES = tuple(round(1 - i * 0.02, 2) for i in range(1, 26))

#: (input de `0. Supuestos`, valor del ensayo, [(celda observada, sentido)]).
#: El valor del ensayo se calcula desde el del caso base cuando hace falta.
DIRECCIONES = (
    ('cubiertos', lambda v: (v or 55) * 1.2,
     (('ingresos', 'sube'), ('rai', 'sube'), ('r_personal', 'baja'))),
    ('ticket_sup', lambda v: (v or 17.0) * 1.2,
     (('ingresos', 'sube'), ('be_dia', 'baja'), ('r_alquiler', 'baja'))),
    ('dias_sup', lambda v: (v or 310) - 30,
     (('ingresos', 'baja'), ('be_dia', 'sube'))),
    ('coste_comida_sup', lambda v: (v or 0.30) + 0.05,
     (('cv_comida', 'sube'), ('r_cogs', 'sube'), ('rai', 'baja'))),
    ('delivery_sup', lambda v: 0.25,
     (('deliv_pyg', 'sube'), ('rai', 'baja'))),
    ('alquiler_sup', lambda v: (v or 2900) + 500,
     (('r_alquiler', 'sube'), ('rai', 'baja'), ('fondo_inv', 'sube'))),
    ('ss_sup', lambda v: (v or 0.33) + 0.05,
     (('personal_pyg', 'sube'), ('personal_hoja', 'sube'), ('rai', 'baja'))),
    ('prestamo_sup', lambda v: (v or 110000) + 40000,
     (('int1', 'sube'), ('cuota', 'sube'), ('dif', 'sube'), ('rai', 'baja'))),
    # B15 / DEMO-2119 — el objetivo era «Devolución de principal del año 2»,
    # que con dos años de CARENCIA vale 0 por diseño: la comprobación no
    # podía pasar nunca y el motor prohibía de facto cualquier carencia >= 2,
    # que la SPEC §2.8 contempla explícitamente. Ahora se resuelve el PRIMER
    # año que amortiza (carencia + 1).
    ('plazo_sup', lambda v: (v or 7) + 3,
     (('cuota', 'baja'), ('cap_amortiza', 'baja'))),
    ('meses_fondo_sup', lambda v: (v or 3) + 3,
     (('fondo_inv', 'sube'), ('caja', 'sube'), ('saldo_min', 'sube'),
      ('rai', 'igual'))),
    ('vida_obra_sup', lambda v: (v or 10) + 10,
     (('amort', 'baja'), ('rai', 'sube'))),
    ('crec2_sup', lambda v: (v or 0.10) + 0.10,
     (('ingresos2', 'sube'), ('ingresos', 'igual'))),
    ('ipc_sup', lambda v: 0.05,
     (('tcf2', 'sube'), ('rai', 'igual'))),
)


def demos(carpeta, demos_dir, pid, origen=None):
    """Las 8 demostraciones de §2.11, evaluadas con pycel."""
    res = {'fallos': [], 'demostraciones_2_11': {}}
    path = _pf(carpeta)
    if path is None:
        return res
    # RT-20 — `post()` filtra por molde y `demos()` no: al correr un producto
    # de línea B, grupo_a intentaba demostrar sus fórmulas sobre un libro
    # B-γ/B-δ, no encontraba las hojas de línea A y metía un FALLO en el
    # informe. Los dos productos de línea B salían con exit 1 por un fallo
    # que no era del producto, sino del alcance del módulo, y eso enmascara
    # los fallos reales cuando llegue grupo_b.
    import openpyxl as _px0
    try:
        _det = motor.detectar(_px0.load_workbook(path), os.path.basename(path))
    except motor.MoldeDesconocido:
        _det = {'molde': None}
    if _det.get('molde') not in MOLDES:
        res['fuera_de_alcance'] = {
            'fichero': os.path.basename(path), 'molde': _det.get('molde'),
            'motivo': 'grupo_a sólo demuestra los moldes ' + str(MOLDES)}
        return res
    os.makedirs(demos_dir, exist_ok=True)
    wb, mapa = _mapa(path)
    nombres = dict((motor.norm(ws.title), ws.title) for ws in wb.worksheets)

    _voc = {}          # lo rellena el bloque M9, más abajo (ver comentario)

    def R(hoja_, rotulo, col='B'):
        clave = motor.norm(hoja_)
        rot = traducir(rotulo, _voc.get('rx'), _voc.get('mapa') or {})
        fila = mapa.get(clave, {}).get(motor.norm(rot))
        if fila is None:
            return None
        return "'" + nombres[clave] + "'!" + col + str(fila)

    # A6 — los rótulos de los cinco ratios los puede renombrar el módulo de
    # contenido (`UMBRALES[i][1]`): «Alquiler / Ventas» es «Aparcamiento /
    # Ventas» en un producto sin local. Las demostraciones resuelven por
    # RÓTULO, así que tienen que preguntar por el del molde o dan un falso
    # «no se localiza la fila» y tumban el dry-run entero.
    try:
        import importlib as _il
        _cont = _il.import_module('contenido_' + pid.replace('-', '_') + '.a')
    except Exception:                                        # noqa: BLE001
        _cont = None
    _rot_ratio = dict((u[0], u[1])
                      for u in (getattr(_cont, 'UMBRALES', None) or ())
                      if len(u) > 1 and u[1])
    # ⚠️ M9 — las demostraciones localizan las filas POR RÓTULO, y el
    # vocabulario del oficio acaba de renombrarlas («Cubiertos/día» →
    # «Transacciones/día»). Sin traducir el rótulo que se busca, `demos()`
    # canta «no se localizan las filas be_dia, cubiertos» y tumba el dry-run
    # de los tres hermanos que lo usan. Medido en la panadería, 2026-09-05.
    _voc['rx'], _voc['mapa'] = compilar_vocabulario(
        getattr(_cont, 'VOCABULARIO', None) if _cont is not None else None)

    def RATIO(clave, defecto):
        return _rot_ratio.get(clave, defecto)

    # ⚠️ M5 / R22-TAP-07 — se mide sobre el CHECKLIST GUARDADO, no sobre el
    # workbook en memoria: es la única forma de que el gate vea lo mismo que
    # el comprador. Un reemplazo que no casa no rompe nada y no cambia ningún
    # contador; desde aquí, tumba el dry-run.
    _ck = None
    for _n in sorted(os.listdir(carpeta)):
        if _n.startswith('checklist') and _n.endswith('.xlsx'):
            _ck = os.path.join(carpeta, _n)
            break
    _reem = ((getattr(_cont, 'CHECKLIST', None) or {}).get('reemplazos')
             if _cont is not None else None)
    if _ck and _reem:
        _no, _mu = auditar_reemplazos(_px0.load_workbook(_ck), _reem)
        res['checklist_reemplazos'] = {
            'fichero': os.path.basename(_ck), 'declarados': len(_reem),
            'no_entregados': [{'patron': p_, 'texto_que_falta': n_,
                               'celda_alcanzada': c_} for p_, n_, c_ in _no],
            'patrones_muertos': [{'patron': p_, 'texto_que_falta': n_}
                                 for p_, n_, _c in _mu]}
        for p_, n_, c_ in _no:
            res['fallos'].append(
                'checklist: el reemplazo ' + repr(p_) + ' de '
                + os.path.basename(_ck) + ' alcanza la celda «'
                + str(c_)[:60] + '» y NO se aplicó: el entregable sale SIN '
                '«' + n_[:80] + '» (M5 / R22-TAP-07)')

    sup = [k for k in nombres if k.startswith('0. supuestos')]
    sup = nombres[sup[0]] if sup else motor.HOJA_SUPUESTOS
    pyg = [k for k in nombres if 'p&l' in k or k.startswith('pyg')]
    pyg = nombres[pyg[0]] if pyg else None
    eq = [k for k in nombres if 'equilibrio' in k]
    eq = nombres[eq[0]] if eq else None
    esc = [k for k in nombres if 'escenario' in k]
    esc = nombres[esc[0]] if esc else None
    tes = [k for k in nombres if 'tesorer' in k]
    tes = nombres[tes[0]] if tes else None
    fin = [k for k in nombres if 'financiaci' in k]
    fin = nombres[fin[0]] if fin else None
    per = [k for k in nombres if k.endswith('personal')]
    per = nombres[per[0]] if per else None
    inv = [k for k in nombres if 'inversi' in k]
    inv = nombres[inv[0]] if inv else None
    if not all((pyg, eq, esc, tes, fin, per, inv)):
        res['fallos'].append('grupo_a: faltan hojas en ' + path)
        return res

    celdas = {
        'cubiertos': R(sup, 'Cubiertos/día (media del año 1)'),
        'delivery': R(sup, 'Ventas por delivery sobre el total'),
        'plazo': R(sup, 'Plazo del préstamo (años)'),
        'carencia': R(sup, 'Carencia de principal (años)'),
        'ingresos': R(pyg, 'INGRESOS TOTALES (sin IVA)'),
        'rai': R(pyg, 'RESULTADO ANTES DE IMPUESTOS'),
        'neto': R(pyg, 'RESULTADO NETO'),
        'personal_pyg': R(pyg, 'Personal (nóminas + Seguridad Social)'),
        'personal_hoja': R(per, 'TOTAL PLANTILLA', 'G'),
        'r_personal': R(pyg, RATIO('r_personal',
                                  'Coste de personal / Ventas')),
        'u_personal': R(pyg, RATIO('u_personal', RATIO(
            'r_personal', 'Coste de personal / Ventas')), 'E'),
        'deliv_pyg': R(pyg, 'Comisiones de delivery'),
        'is1': R(pyg, 'Impuesto de Sociedades'),
        'is2': R(pyg, 'Impuesto de Sociedades', 'C'),
        'is3': R(pyg, 'Impuesto de Sociedades', 'D'),
        'tipo3': R(pyg, 'Tipo de Impuesto de Sociedades aplicado', 'D'),
        'rai2': R(pyg, 'RESULTADO ANTES DE IMPUESTOS', 'C'),
        'bin_ini2': R(pyg, 'Bases negativas pendientes al inicio', 'C'),
        'be_dia': R(eq, 'Cubiertos necesarios al día'),
        'esc_rai': R(esc, 'RESULTADO ANTES DE IMPUESTOS', 'C'),
        'esc_pesimista': R(esc, 'RESULTADO ANTES DE IMPUESTOS', 'B'),
        'esc_optimista': R(esc, 'RESULTADO ANTES DE IMPUESTOS', 'D'),
        'tipo_sup': R(sup, 'Tipo de interés nominal anual'),
        'pagas_sup': R(sup, 'Número de pagas anuales'),
        'vida_maq_sup': R(sup, 'Vida útil de maquinaria y mobiliario (años)'),
        'cuadro_cierre': R(fin, 'Capital pendiente al vencimiento'),
        'esc_neto': R(esc, 'RESULTADO NETO', 'C'),
        'saldo_min': R(tes, 'Saldo mínimo del año'),
        'fondo': R(inv, 'Colchón operativo hasta alcanzar el equilibrio'),
        'cierre': R(fin, 'Capital pendiente al vencimiento'),
        'cv_comida': R(pyg, 'Coste de mercancía — comida'),
        'r_cogs': R(pyg, RATIO('r_cogs',
                               'Coste de mercancía / Ventas')),
        'r_alquiler': R(pyg, RATIO('r_alquiler', 'Alquiler / Ventas')),
        'amort': R(pyg, 'Amortización del inmovilizado'),
        'ingresos2': R(pyg, 'INGRESOS TOTALES (sin IVA)', 'C'),
        'tcf2': R(pyg, 'TOTAL COSTES FIJOS', 'C'),
        'cuota': R(fin, 'Cuota anual durante la amortización'),
        'int1': R(fin, 'Intereses del año 1'),
        'cap2': R(fin, 'Devolución de principal del año 2'),
        'dif': R(fin, 'Diferencia (origen - usos)'),
        'dif_pct': R(fin, 'Diferencia sobre los usos (%)'),
        'holgura_max': R(fin, 'Exceso de financiación admisible (%)'),
        'fondo_inv': R(inv, 'Colchón operativo hasta alcanzar el equilibrio'),
        'caja': R(inv, 'NECESIDAD TOTAL DE CAJA AL ARRANQUE'),
        'iva_inv': R(inv, 'IVA soportado sobre la inversión (recuperable)'),
        'ticket_sup': R(sup, 'Ticket medio SIN IVA (€)'),
        'dias_sup': R(sup, 'Días de apertura al año'),
        'coste_comida_sup': R(
            sup, 'Coste de mercancía sobre las ventas de COMIDA'),
        'alquiler_sup': R(sup, 'Alquiler mensual del local (€)'),
        'ss_sup': R(sup, 'Seguridad Social a cargo de la empresa'),
        'prestamo_sup': R(sup, 'Préstamo bancario solicitado (€)'),
        'plazo_sup': R(sup, 'Plazo del préstamo (años)'),
        'meses_fondo_sup': R(sup, 'Fondo de maniobra (meses de costes fijos)'),
        'vida_obra_sup': R(sup, 'Vida útil de obra e instalaciones (años)'),
        'crec2_sup': R(sup, 'Crecimiento de volumen del año 2'),
        'ipc_sup': R(sup, 'Subida anual de los costes fijos'),
        'delivery_sup': R(sup, 'Ventas por delivery sobre el total'),
        'y4': R(fin, None) if False else None,
    }
    # B15 — el primer año que AMORTIZA de verdad: con carencia 1 es el 2 (lo
    # de siempre), con carencia 2 el 3, y con una carencia mayor que la
    # proyección la comprobación se omite en vez de fallar.
    _car = 1
    if celdas.get('carencia'):
        _v = wb[sup][celdas['carencia'].split('!')[-1]].value
        if isinstance(_v, (int, float)) and not isinstance(_v, bool):
            _car = int(_v)
    _n_amort = _car + 1
    celdas['cap_amortiza'] = (
        R(fin, 'Devolución de principal del año ' + str(_n_amort))
        if 1 <= _n_amort <= 3 else None)
    res['primer_ano_que_amortiza'] = {'carencia': _car, 'ano': _n_amort,
                                      'celda': celdas['cap_amortiza']}
    faltan = [k for k, v in celdas.items()
              if v is None and k not in ('y4', 'cap_amortiza')]
    if faltan:
        res['fallos'].append('grupo_a: no se localizan las filas '
                             + ', '.join(sorted(faltan)))
        return res

    # RC-32 — el mapa de celdas de §4.2 de la SPEC quedó OBSOLETO al
    # reestructurar el libro en T1: seis de las nueve referencias que manda
    # citar apuntan hoy a otra magnitud (`1!B46` es el IVA soportado, no la
    # inversión; `2!B34` son 600 € de PRL, no el resultado neto). Si
    # `guion_<pid>.py` se escribe siguiendo la tabla, el docx publicará el
    # IVA como inversión total y el gate de cifras lo dará por bueno porque
    # coincidirá con la celda citada. Aquí se publica el mapa REAL, leído del
    # fichero por RÓTULO, para que T9 cite celdas y no coordenadas de memoria.
    mapa_t9 = {}
    for magnitud, hoja_, rotulo, columna in (
            ('inversion_total', inv, 'INVERSIÓN TOTAL (suma de los bloques)',
             'B'),
            ('capex', inv, 'CAPEX (inversión sin el fondo de maniobra)', 'B'),
            ('iva_de_la_inversion', inv,
             'IVA soportado sobre la inversión (recuperable)', 'B'),
            ('necesidad_de_caja', inv,
             'NECESIDAD TOTAL DE CAJA AL ARRANQUE', 'B'),
            ('fondo_de_maniobra', inv,
             'Colchón operativo hasta alcanzar el equilibrio', 'B'),
            ('amortizacion_anual', inv,
             'Amortización anual del inmovilizado', 'B'),
            ('facturacion_ano_1', pyg, 'INGRESOS TOTALES (sin IVA)', 'B'),
            ('facturacion_ano_2', pyg, 'INGRESOS TOTALES (sin IVA)', 'C'),
            ('facturacion_ano_3', pyg, 'INGRESOS TOTALES (sin IVA)', 'D'),
            ('margen_bruto', pyg, 'MARGEN BRUTO', 'B'),
            ('resultado_antes_de_impuestos', pyg,
             'RESULTADO ANTES DE IMPUESTOS', 'B'),
            ('resultado_neto_ano_1', pyg, 'RESULTADO NETO', 'B'),
            ('resultado_neto_ano_3', pyg, 'RESULTADO NETO', 'D'),
            ('margen_neto_pct', pyg,
             RATIO('r_neto', 'Resultado neto / Ventas'), 'B'),
            ('ratio_personal_pct', pyg,
             RATIO('r_personal', 'Coste de personal / Ventas'), 'B'),
            ('ratio_cogs_pct', pyg,
             RATIO('r_cogs', 'Coste de mercancía / Ventas'), 'B'),
            ('coste_de_personal', pyg,
             'Personal (nóminas + Seguridad Social)', 'B'),
            ('plantilla_personas', per, 'TOTAL PLANTILLA', 'B'),
            ('plantilla_jornadas', per, 'TOTAL PLANTILLA', 'C'),
            ('coste_plantilla_ano', per, 'TOTAL PLANTILLA', 'G'),
            ('break_even_contable_dia', eq, 'Cubiertos necesarios al día',
             'B'),
            ('break_even_caja_dia', eq, 'Cubiertos necesarios al día (caja)',
             'B'),
            ('break_even_ingresos', eq, 'Ingresos necesarios al año (caja)',
             'B'),
            ('saldo_minimo_de_caja', tes, 'Saldo mínimo del año', 'B'),
            ('mes_de_caja_minima', tes, 'Mes en el que la caja toca fondo',
             'B'),
            ('payback_anos', tes, 'Payback del proyecto (años)', 'B'),
            ('cuota_anual_prestamo', fin,
             'Cuota anual durante la amortización', 'B'),
            ('dscr_minimo', fin, 'DSCR mínimo de todo el cuadro', 'B'),
            ('origen_menos_usos', fin, 'Diferencia (origen - usos)', 'B')):
        mapa_t9[magnitud] = R(hoja_, rotulo, columna)
    res['mapa_de_celdas_para_T9'] = {
        'fichero': os.path.basename(path),
        'aviso': 'Mapa REAL leído del fichero por rótulo. La tabla de §4.2 de '
                 'la SPEC está obsoleta desde T1 (RC-32): el guion tiene que '
                 'resolver por RÓTULO, no por coordenada, y abortar si el '
                 'rótulo no aparece.',
        'celdas': mapa_t9,
        'sin_localizar': sorted(k for k, v in mapa_t9.items() if v is None)}

    # ---- A7 (gate propio) — el rango del SUMIF del IVA soportado de la
    # inversión tiene que cubrir EXACTAMENTE el mismo tramo que los SUM de
    # los subtotales de bloque. Es el defecto REF-14: el SUMIF se quedaba una
    # fila corta y dejaba fuera los imprevistos, que sí están marcados «Sí».
    ws_inv = wb[inv]
    rx_sumif = re.compile(r'SUMIF\(\$E\$(\d+):\$E\$(\d+)')
    rx_sum_b = re.compile(r'SUM\(B(\d+):B(\d+)\)')
    _a = _b = None
    _tramos = []
    for _r in range(1, ws_inv.max_row + 1):
        _v = ws_inv.cell(row=_r, column=2).value
        if not isinstance(_v, str):
            continue
        _rot = motor._rotulo_de_fila(ws_inv, _r, max_col=1) or ''
        _m = rx_sumif.search(_v)
        if _m and 'iva soportado' in motor.norm(_rot):
            _a, _b = int(_m.group(1)), int(_m.group(2))
        _m2 = rx_sum_b.search(_v)
        if _m2 and _es_mayusculas(_rot or ' '):
            _tramos.append((int(_m2.group(1)), int(_m2.group(2))))
    if _a is not None and _tramos:
        _ini = min(t[0] for t in _tramos)
        _fin = max(t[1] for t in _tramos)
        _ok = (_a == _ini and _b == _fin)
        res['gate_A7_sumif_iva_inversion'] = {
            'sumif': 'E%d:E%d' % (_a, _b),
            'subtotales_de_bloque': 'B%d:B%d' % (_ini, _fin), 'ok': _ok}
        if not _ok:
            res['fallos'].append(
                'A7 / REF-14: el SUMIF del IVA soportado de la inversión '
                'cubre E%d:E%d y los subtotales de bloque suman B%d:B%d — '
                'hay partidas marcadas «Sí» fuera del SUMIF' % (_a, _b, _ini,
                                                                _fin))

    xl = _pycel(path)
    base = dict((k, _ev(xl, v)) for k, v in celdas.items() if v)
    res['mapa_de_celdas_para_T9']['valores'] = dict(
        (k, (round(_ev(xl, v), 4) if isinstance(_ev(xl, v), float)
             else _ev(xl, v)))
        for k, v in mapa_t9.items() if v)
    res['demostraciones_2_11']['caso_base'] = dict(
        (k, (round(v, 2) if isinstance(v, float) else v))
        for k, v in base.items())

    def num(v):
        return v if isinstance(v, (int, float)) and not isinstance(v, bool) \
            else None

    # 1 — mover los cubiertos mueve las cuatro cifras estrella
    p1 = _clon(path, os.path.join(demos_dir, 'd1.xlsx'),
               [(sup, celdas['cubiertos'].split('!')[-1],
                 (num(base['cubiertos']) or 55) * 1.4)])
    x1 = _pycel(p1)
    mov = {}
    # ⚠️ `be_dia` YA NO entra aquí: desde RT-05 el punto de equilibrio NO
    # puede moverse con el volumen previsto (un break-even lineal no lo
    # hace), y se comprueba justo debajo con el signo contrario.
    for clave in ('ingresos', 'rai', 'esc_rai'):
        mov[clave] = [base[clave], _ev(x1, celdas[clave])]
    quietas = [k for k, v in mov.items()
               if num(v[0]) is not None and num(v[1]) is not None
               and abs(v[0] - v[1]) < 0.01]
    res['demostraciones_2_11']['1_cubiertos_mueven_el_libro'] = {
        'celda': celdas['cubiertos'], 'valor_nuevo': round(
            (num(base['cubiertos']) or 55) * 1.4, 2),
        'movimiento': dict((k, [round(v[0], 2) if num(v[0]) is not None
                                else v[0],
                                round(v[1], 2) if num(v[1]) is not None
                                else v[1]]) for k, v in mov.items()),
        'ok': not quietas}
    if quietas:
        res['fallos'].append('§2.11.1: al subir los cubiertos NO se mueven '
                             + ', '.join(quietas))

    # 1-bis (RT-05) — el punto de equilibrio NO se mueve con el volumen. Un
    # coste proporcional a las ventas metido en el numerador de costes fijos
    # hacía que la cifra estrella del producto dependiera de un supuesto que
    # no debería tocarla: con 80 cubiertos daba 67,44 y con 110, 68,29.
    be_alto = _ev(x1, celdas['be_dia'])
    p1b = _clon(path, os.path.join(demos_dir, 'd1b.xlsx'),
                [(sup, celdas['cubiertos'].split('!')[-1],
                  max(1, int(round((num(base['cubiertos']) or 55) * 0.4))))])
    be_bajo = _ev(_pycel(p1b), celdas['be_dia'])
    valores_be = [v for v in (num(base['be_dia']), num(be_alto),
                              num(be_bajo)) if v is not None]
    ok1b = len(valores_be) == 3 and (max(valores_be) - min(valores_be)) <= 0.01
    res['demostraciones_2_11']['1bis_break_even_invariante_al_volumen'] = {
        'celda': celdas['be_dia'], 'caso_base': num(base['be_dia']),
        'con_mas_volumen': num(be_alto), 'con_menos_volumen': num(be_bajo),
        'ok': ok1b}
    if not ok1b:
        res['fallos'].append('§2.11.1-bis / RT-05: el punto de equilibrio se '
                             'MUEVE con el volumen previsto ('
                             + str(valores_be) + ')')

    # 2-bis (RT-04) — los escenarios extremos son independientes del
    # realista: mover un input que sólo pertenece al realista no puede
    # cambiar el resultado del pesimista ni el del optimista.
    esc_pes_1 = _ev(x1, celdas['esc_pesimista'])
    esc_opt_1 = _ev(x1, celdas['esc_optimista'])
    contaminados = []
    for clave, antes, ahora in (('pesimista', base['esc_pesimista'],
                                 esc_pes_1),
                                ('optimista', base['esc_optimista'],
                                 esc_opt_1)):
        if num(antes) is None or num(ahora) is None \
                or abs(antes - ahora) > 0.01:
            contaminados.append([clave, antes, ahora])
    res['demostraciones_2_11']['2bis_escenarios_independientes'] = {
        'input_movido': celdas['cubiertos'],
        'pesimista': [base['esc_pesimista'], esc_pes_1],
        'optimista': [base['esc_optimista'], esc_opt_1],
        'ok': not contaminados}
    if contaminados:
        res['fallos'].append('§2.11.2-bis / RT-04: mover un input del '
                             'REALISTA cambia el resultado de '
                             + ', '.join(c[0] for c in contaminados)
                             + ' ' + str(contaminados))

    # 2 — el escenario Realista reproduce el P&L al céntimo
    d2 = None
    if num(base['esc_rai']) is not None and num(base['rai']) is not None:
        d2 = round(abs(base['esc_rai'] - base['rai']), 4)
    res['demostraciones_2_11']['2_realista_igual_al_pyg'] = {
        'escenarios_realista': base['esc_rai'], 'pyg_ano_1': base['rai'],
        'diferencia': d2, 'ok': d2 is not None and d2 <= 0.01}
    if d2 is None or d2 > 0.01:
        res['fallos'].append('§2.11.2: «Realista» y el año 1 del P&L difieren '
                             + str(d2))

    # 3 — el P&L lee el personal de su hoja y el ratio cumple su umbral
    d3 = None
    if num(base['personal_pyg']) is not None \
            and num(base['personal_hoja']) is not None:
        d3 = round(abs(base['personal_pyg'] - base['personal_hoja']), 4)
    ratio_ok = (num(base['r_personal']) is not None
                and num(base['u_personal']) is not None
                and base['r_personal'] <= base['u_personal'] + 1e-9)
    res['demostraciones_2_11']['3_personal_del_pyg_es_el_de_su_hoja'] = {
        'pyg': base['personal_pyg'], 'hoja_personal': base['personal_hoja'],
        'diferencia': d3, 'ratio': base['r_personal'],
        'umbral': base['u_personal'], 'ok': (d3 == 0) and ratio_ok}
    if d3 != 0:
        res['fallos'].append('§2.11.3: el P&L no imputa el coste de la hoja '
                             'Personal (diferencia ' + str(d3) + ')')
    if not ratio_ok:
        res['fallos'].append('§2.11.3 / DOM-13: el caso base NO pasa su '
                             'propio semáforo de personal ('
                             + str(base['r_personal']) + ' > '
                             + str(base['u_personal']) + ')')

    # 4 — delivery a cero deja la línea a cero
    p4 = _clon(path, os.path.join(demos_dir, 'd4.xlsx'),
               [(sup, celdas['delivery'].split('!')[-1], 0)])
    x4 = _pycel(p4)
    v4 = _ev(x4, celdas['deliv_pyg'])
    res['demostraciones_2_11']['4_delivery_a_cero'] = {
        'linea_delivery': v4, 'ok': num(v4) == 0}
    if num(v4) != 0:
        res['fallos'].append('§2.11.4: con delivery al 0 % la línea vale '
                             + str(v4))

    # 5 — IS: pérdida en el año 1, BIN compensada en el 2, tipo reducido
    # El escenario se elige para que se vean las TRES ramas: año 1 en
    # pérdidas (impuesto cero), año 2 con base positiva MENOR que la base
    # negativa pendiente (impuesto cero por compensación) y año 3 tributando
    # al tipo de entidad de nueva creación, que es su primer ejercicio con
    # base positiva.
    base_cub = num(base['cubiertos']) or DEMO5_CUBIERTOS
    ensayos5 = []
    ok5 = False
    cub5 = DEMO5_CUBIERTOS
    is1 = is2 = is3 = rai1 = rai2 = tipo3 = bin2 = None
    for factor in DEMO5_FACTORES:
        cub5 = max(1, int(round(base_cub * factor)))
        p5 = _clon(path, os.path.join(demos_dir, 'd5.xlsx'),
                   [(sup, celdas['cubiertos'].split('!')[-1], cub5)])
        x5 = _pycel(p5)
        is1, is2, is3 = (_ev(x5, celdas['is1']), _ev(x5, celdas['is2']),
                         _ev(x5, celdas['is3']))
        rai1, rai2 = _ev(x5, celdas['rai']), _ev(x5, celdas['rai2'])
        tipo3 = _ev(x5, celdas['tipo3'])
        bin2 = _ev(x5, celdas['bin_ini2'])
        ok5 = (num(rai1) is not None and rai1 < 0 and num(is1) == 0
               and num(rai2) is not None and rai2 > 0 and num(is2) == 0
               and num(bin2) is not None and bin2 > rai2
               and num(tipo3) is not None and abs(tipo3 - 0.15) < 1e-9
               and num(is3) is not None and is3 > 0)
        ensayos5.append({'cubiertos': cub5, 'rai_ano_1': rai1,
                         'rai_ano_2': rai2, 'bin_ano_2': bin2, 'ok': ok5})
        if ok5:
            break
    res['demostraciones_2_11']['5_IS_con_BIN_y_tipo_de_nueva_creacion'] = {
        'cubiertos_del_ensayo': cub5, 'ensayos': ensayos5,
        'rai_ano_1': rai1, 'is_ano_1': is1,
        'rai_ano_2': rai2, 'bin_pendiente_ano_2': bin2, 'is_ano_2': is2,
        'is_ano_3': is3, 'tipo_ano_3': tipo3, 'ok': ok5}
    if not ok5:
        res['fallos'].append('§2.11.5: no se demuestran las tres ramas del '
                             'impuesto (RAI1 ' + str(rai1) + ', IS1 '
                             + str(is1) + ', RAI2 ' + str(rai2) + ', BIN2 '
                             + str(bin2) + ', IS2 ' + str(is2) + ', tipo3 '
                             + str(tipo3) + ', IS3 ' + str(is3) + ')')

    # 6 — la caja nunca se agota con el fondo dotado
    ok6 = num(base['saldo_min']) is not None and base['saldo_min'] >= 0
    res['demostraciones_2_11']['6_tesoreria_sin_saldo_negativo'] = {
        'saldo_minimo': base['saldo_min'], 'fondo': base['fondo'], 'ok': ok6}
    if not ok6:
        res['fallos'].append('§2.11.6: el saldo mínimo de tesorería es '
                             + str(base['saldo_min'])
                             + ': hay que subir el fondo de maniobra')

    # 7 — carencia >= plazo: se anula y el cuadro cierra en cero
    p7 = _clon(path, os.path.join(demos_dir, 'd7.xlsx'),
               [(sup, celdas['plazo'].split('!')[-1], 3),
                (sup, celdas['carencia'].split('!')[-1], 3)])
    x7 = _pycel(p7)
    cierre = _ev(x7, celdas['cierre'])
    # las filas del cuadro llevan el AÑO como número en la columna A, así que
    # no tienen rótulo de texto y no se pueden buscar por él
    apagados = []
    import openpyxl as _px
    ws7 = motor.hoja(_px.load_workbook(p7), fin, obligatoria=True)
    for anio in (4, 5):
        for r in range(1, ws7.max_row + 1):
            if ws7.cell(row=r, column=1).value == anio:
                for col in ('C', 'D', 'E'):
                    apagados.append([anio, col,
                                     _ev(x7, "'" + fin + "'!" + col + str(r))])
                break
    ok7 = (num(cierre) is not None and abs(cierre) <= 0.5
           and len(apagados) == 6
           and all(v in ('', None) or num(v) == 0
                   for _a, _c, v in apagados))
    res['demostraciones_2_11']['7_carencia_igual_al_plazo'] = {
        'capital_al_vencimiento': cierre, 'cuotas_anos_4_y_5': apagados,
        'ok': ok7}
    if not ok7:
        res['fallos'].append('§2.11.7: con plazo 3 y carencia 3 el cuadro no '
                             'cierra en cero o los años 4-5 siguen vivos ('
                             + str(cierre) + ', ' + str(apagados) + ')')

    # 8 — libro EN BLANCO: ni un semáforo verde ni un 0,0 % falso
    # se vacían TODAS las celdas de entrada de la hoja de supuestos, no una:
    # el libro en blanco es el estado en el que un semáforo verde o un
    # «0,0 %» mienten más caro
    # ⚠️ se vacían las verdes de TODAS las hojas, no sólo las de Supuestos:
    # los escenarios extremos y el reparto mensual tienen sus propios inputs,
    # y dejarlos vivos hacía que el «libro en blanco» siguiera imprimiendo
    # cifras (RT-18).
    import openpyxl as _px2
    wb8 = _px2.load_workbook(path)
    blancos = [(ws8.title, c.coordinate, None) for ws8 in wb8.worksheets
               for row in ws8.iter_rows() for c in row
               if motor.es_verde(c)
               and not motor._es_formula(c.value)]
    p8 = _clon(path, os.path.join(demos_dir, 'd8.xlsx'), blancos)
    x8 = _pycel(p8)
    vacias = {}
    for clave in ('ingresos', 'rai', 'neto', 'r_personal', 'be_dia',
                  'esc_rai', 'saldo_min', 'cierre'):
        vacias[clave] = _ev(x8, celdas[clave])
    falsos = [k for k, v in vacias.items()
              if isinstance(v, (int, float)) and not isinstance(v, bool)]
    res['demostraciones_2_11']['8_libro_en_blanco_sin_falsos_verdes'] = {
        'valores': vacias, 'ok': not falsos}
    if falsos:
        res['fallos'].append('§2.11.8: con el libro en blanco siguen dando '
                             'número (y por tanto semáforo) ' + ', '.join(
                                 falsos))
    # 9 — CADA input nuevo mueve lo que debe, y en la dirección que debe.
    # No basta con que «se mueva»: una amortización que sube al alargar la
    # vida útil o un break-even que sube al subir el ticket son fórmulas mal
    # puestas que ningún gate de formato detecta.
    direcciones = []
    for clave, nuevo_valor, esperado in DIRECCIONES:
        celda = celdas.get(clave)
        if celda is None:
            res['fallos'].append('§2.11.9: no se localiza el input ' + clave)
            continue
        valor = nuevo_valor(num(base.get(clave)) if clave in base else None)
        ruta = os.path.join(demos_dir, 'd9-' + clave + '.xlsx')
        xln = _pycel(_clon(path, ruta,
                           [(sup, celda.split('!')[-1], valor)]))
        for objetivo, sentido in esperado:
            if celdas.get(objetivo) is None:
                direcciones.append({'input': clave, 'objetivo': objetivo,
                                    'omitido': 'no aplica en este producto '
                                               '(carencia >= 3 años: ningún '
                                               'año proyectado amortiza)',
                                    'ok': True})
                continue
            antes = num(base.get(objetivo))
            if antes is None:
                antes = num(_ev(xl, celdas[objetivo]))
            ahora = num(_ev(xln, celdas[objetivo]))
            if antes is None or ahora is None:
                ok = False
                delta = None
            else:
                delta = round(ahora - antes, 4)
                ok = (delta > 0.005 if sentido == 'sube'
                      else delta < -0.005 if sentido == 'baja'
                      else abs(delta) <= 0.005)
            direcciones.append({'input': clave, 'celda_input': celda,
                                'valor_nuevo': valor, 'objetivo': objetivo,
                                'celda_objetivo': celdas[objetivo],
                                'sentido_esperado': sentido,
                                'antes': antes, 'despues': ahora,
                                'delta': delta, 'ok': ok})
            if not ok:
                res['fallos'].append(
                    '§2.11.9: cambiar ' + clave + ' (' + celda + ') NO hace '
                    'que ' + objetivo + ' (' + celdas[objetivo] + ') '
                    + sentido + ': ' + str(antes) + ' → ' + str(ahora))
    res['demostraciones_2_11']['9_direccion_de_cada_calculo'] = {
        'comprobaciones': direcciones,
        'ok': all(d['ok'] for d in direcciones)}

    # 10 (RT-06) — tipo de interés al 0 %: un préstamo familiar, un ENISA o
    # un ICO bonificado. La anualidad algebraica dividía entre cero, el
    # cuadro se apagaba entero y «Capital pendiente al vencimiento»
    # certificaba EN VERDE que se habían devuelto 110.000 € que nadie había
    # devuelto.
    p10 = _clon(path, os.path.join(demos_dir, 'd10.xlsx'),
                [(sup, celdas['tipo_sup'].split('!')[-1], 0)])
    x10 = _pycel(p10)
    cuota0 = _ev(x10, celdas['cuota'])
    cierre0 = _ev(x10, celdas['cuadro_cierre'])
    int0 = _ev(x10, celdas['int1'])
    ok10 = (num(cuota0) is not None and cuota0 > 0
            and num(cierre0) is not None and abs(cierre0) <= 0.5
            and num(int0) is not None and abs(int0) <= 0.01)
    res['demostraciones_2_11']['10_tipo_al_cero_por_ciento'] = {
        'cuota': cuota0, 'capital_al_vencimiento': cierre0,
        'intereses_ano_1': int0, 'ok': ok10}
    if not ok10:
        res['fallos'].append('§2.11.10 / RT-06: con el tipo al 0 % la cuota '
                             'es ' + str(cuota0) + ' y el capital pendiente '
                             'al vencimiento ' + str(cierre0))

    # 11 (RT-07) — plazo largo: una línea ICO para inversión en hostelería va
    # a 12-15 años. Con el cuadro de 10 filas quedaban 25.570,74 € sin
    # amortizar y la hoja acusaba al usuario de un error del generador.
    p11 = _clon(path, os.path.join(demos_dir, 'd11.xlsx'),
                [(sup, celdas['plazo_sup'].split('!')[-1], motor.PLAZO_MAX)])
    cierre11 = _ev(_pycel(p11), celdas['cuadro_cierre'])
    ok11 = num(cierre11) is not None and abs(cierre11) <= 0.5
    res['demostraciones_2_11']['11_plazo_maximo_cierra_el_cuadro'] = {
        'plazo': motor.PLAZO_MAX, 'capital_al_vencimiento': cierre11,
        'ok': ok11}
    if not ok11:
        res['fallos'].append('§2.11.11 / RT-07: con plazo '
                             + str(motor.PLAZO_MAX) + ' quedan '
                             + str(cierre11) + ' € sin amortizar')

    # 12 (RT-08) — robustez: un 0 en un DIVISOR no puede mejorar el resultado
    # ni dejar el coste de personal a cero sin rastro visible.
    robustez = []
    for clave, objetivo, regla in (
            ('vida_obra_sup', 'rai', 'no_mejora'),
            ('vida_maq_sup', 'rai', 'no_mejora'),
            ('pagas_sup', 'personal_pyg', 'no_cero'),
            ('dias_sup', 'ingresos', 'vacio'),
            ('cubiertos', 'ingresos', 'vacio'),
            ('meses_fondo_sup', 'fondo', 'vacio'),
            ('plazo_sup', 'cuadro_cierre', 'rastro')):
        celda = celdas.get(clave)
        if celda is None:
            continue
        ruta = os.path.join(demos_dir, 'd12-' + clave + '.xlsx')
        xz = _pycel(_clon(path, ruta, [(sup, celda.split('!')[-1], 0)]))
        v = _ev(xz, celdas[objetivo])
        antes = num(base.get(objetivo))
        if regla == 'no_mejora':
            ok = num(v) is None or antes is None or v <= antes + 0.01
        elif regla == 'no_cero':
            ok = num(v) is not None and abs(v) > 0.01
        elif regla == 'vacio':
            ok = num(v) is None
        else:                                    # 'rastro'
            ok = num(v) is not None and abs(v) > 0.5
        robustez.append({'input': clave, 'celda': celda, 'objetivo': objetivo,
                         'regla': regla, 'antes': antes, 'con_cero': v,
                         'ok': ok})
        if not ok:
            res['fallos'].append('§2.11.12 / RT-08: con ' + clave + ' a 0, '
                                 + objetivo + ' pasa de ' + str(antes)
                                 + ' a ' + str(v) + ' (regla ' + regla + ')')
    res['demostraciones_2_11']['12_robustez_de_los_divisores'] = {
        'comprobaciones': robustez, 'ok': all(r['ok'] for r in robustez)}

    # 12-bis (RD-34) — el origen de fondos tiene que CUBRIR los usos, y sin
    # pasarse: el semáforo sólo comprobaba que la diferencia no fuese
    # negativa, así que un exceso de 20.000 € pasaba igual de verde.
    dif = num(base.get('dif'))
    dif_pct = num(base.get('dif_pct'))
    tope = num(base.get('holgura_max'))
    ok12b = (dif is not None and dif >= -0.01
             and dif_pct is not None and tope is not None
             and dif_pct <= tope + 1e-9)
    res['demostraciones_2_11']['12bis_origen_cubre_los_usos'] = {
        'diferencia': dif, 'diferencia_pct': dif_pct, 'tope': tope,
        'ok': ok12b}
    if not ok12b:
        res['fallos'].append('§2.11.12-bis / RD-34: el origen de fondos no '
                             'cuadra con los usos (diferencia ' + str(dif)
                             + ' €, ' + str(dif_pct) + ' sobre los usos, '
                             'tope ' + str(tope) + ')')

    # 13 (RT-18) — libro EN BLANCO: además de no dar errores, ninguna celda
    # puede imprimir «0 €» ni «0,0 %». Un libro vacío entregado así parece un
    # plan con cifras.
    ceros = []
    import openpyxl as _px3
    wb13 = _px3.load_workbook(p8)
    for ws13 in wb13.worksheets:
        for row13 in ws13.iter_rows():
            for c13 in row13:
                if not motor._es_formula(c13.value):
                    continue
                fmt13 = c13.number_format or ''
                if '\u20ac' not in fmt13 and '%' not in fmt13:
                    continue
                v13 = _ev(x8, "'" + ws13.title + "'!" + c13.coordinate)
                if isinstance(v13, (int, float)) \
                        and not isinstance(v13, bool) and v13 == 0:
                    ceros.append(ws13.title + '!' + c13.coordinate)
    res['demostraciones_2_11']['13_libro_en_blanco_sin_ceros'] = {
        'celdas_con_cero': ceros[:40], 'total': len(ceros),
        'ok': not ceros}
    if ceros:
        res['fallos'].append('§2.11.13 / RT-18: con el libro en blanco '
                             + str(len(ceros)) + ' celdas imprimen 0 € o '
                             '0,0 % (' + ', '.join(ceros[:8]) + ')')
    return res
