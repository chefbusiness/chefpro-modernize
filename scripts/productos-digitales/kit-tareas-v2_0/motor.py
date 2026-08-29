#!/usr/bin/env python3
"""
motor.py — Motor COMÚN de la familia «▸» de kits de tareas (v2.0).

Implementa §1 y §2 de `scripts/productos-digitales/kit-tareas-v2-SPEC.md`: las
correcciones ESTRUCTURALES que se repiten en los 12 kits con 08/09 de negocio y
caja (kit-tareas, cafetería, pizzería, hamburguesería, dark-kitchen, bar,
catering, chocolatería, heladería, hotel 18/19, restaurante-creativo 10/11) y en
los checklists «▸» 01-07 + BONUS.

REGLA DE ORO: **todo se detecta por CABECERA, nunca por nombre de fichero.**
Una hoja entra al motor sólo si su fila de cabecera es una de las cinco de la
familia:

  · checklist  →  fila con A='Nº'|'#', B='Tarea' y una columna «✓ Completada»
                  (los 01-17 del kit de hotel llevan '✓' a secas y fila de
                  cabecera en la 5: NO entran, y por eso quedan intactos)
  · registro mensual de caja  →  fila 'Fecha | Fondo Apertura | Ventas … '
  · recuento de caja          →  fila 'Denominación | Cantidad | Subtotal (€)'
  · calendario ▸              →  fila 'Mes | Fecha / Evento | Tareas Clave | Antelación'
  · briefing ▸                →  B1 con «Briefing» y B2 con «Fecha:»

Un FICHERO se procesa si tiene al menos una hoja reconocida; si no, no se abre
ni su hoja de Instrucciones. Es lo que hace que `--producto kit-tareas-hotel`
toque 18 y 19 y deje los 01-17 byte a byte como estaban.

IDEMPOTENCIA: todo lo que escribe el motor es escritura ABSOLUTA o va detrás de
un centinela estructural (¿existe ya la columna «Firma»? ¿ya hay 5 filas
libres?). DV y CF se VACÍAN al entrar en las hojas reconocidas y se reconstruyen
enteros: si se acumulasen, la 2.ª pasada tendría el doble de reglas y el gate de
idempotencia lo cazaría.

Las utilidades `_traducir_formula`, `_rangos_dv`, `_restaurar_dv`,
`_desplazar_rango`, `insertar_columna`, `insertar_filas` y `print_setup` vienen
de `kit-pasteleria-v2_0-postprocess.py` (líneas 642-760), probado en producción
el 2026-08-22; `eliminar_filas` es su simétrica, que allí no hacía falta.
El bloque de caja (`_fondo_de_caja`, `ajustar_09`) se porta de las líneas
1327-1520 del mismo fichero.
"""
import copy
import difflib
import math
import re
import unicodedata

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Paleta y vocabulario de la familia
# ==========================================================================
VERDE = 'E8F5E9'        # celda editable (la FAQ de la landing lo promete así)
VERDE_OK = 'C8E6C9'     # fila completada (formato condicional)
#: CB-E9 (d) — el mismo verde, un tono más marcado, para la sub-familia CB.
#: `C8E6C9` y el `E8F5E9` de las celdas editables se distinguen mal impresos en
#: gris, justo en las columnas «Zona» y «Responsable», que van verdes: la fila
#: completada no se ve. **NO se cambia el de la familia**: el CF no entra en la
#: comparación de `regresion.py`, así que un cambio global habría repintado los
#: 11 kits publicados sin que el gate bloqueante lo cantase. Queda escrito en
#: el informe como decisión para el orquestador.
VERDE_OK_CB = 'A5D6A7'
#: CB-E9 (a/b) — presupuesto de ancho de una fila en A4 apaisado con
#: `fitToWidth = 1`, en unidades de columna de Excel, y tope de la columna de
#: texto. El molde ▸ de CB mide 5+48+14+20+13+12+16 = 128, así que caben 12
#: unidades más sin salirse.
ANCHO_A4 = 140
ANCHO_TAREA_CB = 60


def verde_ok():
    """CB-E9 (d) — color de la fila completada del kit en curso."""
    return VERDE_OK_CB if sub_cb() else VERDE_OK
SECCION = 'E3F2FD'      # banda de sección (TEC-15: el verde ya NO va aquí)
AMBAR = 'FFF3CD'        # aviso de descuadre
GRIS = 'F5F5F5'
FMT_EUR = '#,##0.00 €'
FMT_ENT = '0'

CAB_MARCA = '✓ Completada'
#: Molde P4 (catering, chocolatería, heladería, hotel, restaurante-creativo):
#: la columna de marca se llama «✓» a secas y la cabecera se REPITE en cada
#: sección. No entra por `geometria`; la trata `normalizar_p4`.
CAB_P4 = '✓'
DV_LISTA = '"✓,—,N/A"'          # ✓ , — , N/A   (§1.5 / TEC-08)
#: DV-R3-C — el mensaje decía «Las marcadas N/A o — salen del total», que
#: CONTRADICE la fórmula vigente desde DOM-R2-02 (sólo N/A sale). La fórmula
#: era correcta; el texto que ve el cliente, no. Afecta a los 11 kits.
DV_ERROR = ('Usa el desplegable: ✓, — o N/A. N/A = no aplica, sale del '
            'total; — = no hecha, cuenta como pendiente.')
DV_ERROR_TIT = 'Marca no válida'
#: Listas de desplegable heredadas que el motor sustituye por `DV_LISTA` en
#: TODO el kit (§R3-e): un mismo producto no puede entregar dos semánticas de
#: conteo en dos ficheros distintos.
def _es_lista_de_marca(f1):
    return isinstance(f1, str) and f1.startswith('"✓') and f1 != DV_LISTA
PIE_VIEJO = '© AI Chef Pro — aichef.pro'
MARCA_OK = '✓'
MARCA_NO = '—'
HOLGURA = 5                                # §2.2 — 5 filas libres EN el rango

ETIQ_FONDO = 'Fondo de caja inicial (€)'
ETIQ_FONDO_RESUMEN = 'Fondo de caja inicial (−)'
ETIQ_EFECTIVO = 'Total Efectivo (recuento)'
#: DOM-R2-01 / TEC-R2-11 — la cifra que el Registro Mensual pide y que el
#: Resumen de Cierre no enseñaba por ninguna parte. Sin esta fila el operario
#: transcribe el recuento BRUTO y el registro marca descuadre = fondo todos los
#: días.
ETIQ_VENTAS_EF = 'Ventas en efectivo (contado − fondo)'
ETIQ_Z = 'Z del TPV'
#: DOM-R2-01 — «Ventas Efectivo» era ambiguo (¿el cajón o el cajón menos el
#: fondo?). El registro pasa a pedir el recuento BRUTO y descuenta el fondo él
#: mismo, igual que hace el Cierre de Caja.
CAB_EFECTIVO = 'Efectivo Contado'

BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, '
       'en cocina desde los 17 años · johnguerrero.es')
RX_BIO = re.compile(r'Diseñado por John Guerrero')
RX_VERSION = re.compile(r'^Versi[óo]n \d+\.\d+ · ')
RX_CONTACTO = re.compile(r'^Contacto:')
RX_COPY = re.compile(r'^©')
RX_HORA = re.compile(r'^([01]?\d|2[0-3]):([0-5]\d)$')
RX_HORA_EN_TEXTO = re.compile(r'\b([01]?\d|2[0-3]):([0-5]\d)\b')

DIAS = ('lunes', 'martes', 'miércoles', 'miercoles', 'jueves', 'viernes',
        'sábado', 'sabado', 'domingo')
RX_CADENCIA = re.compile(
    r'^(mensual|trimestral|anual|semestral|quincenal|semanal|d[ií]a \d|'
    r'1[ºo]? de mes|fin de mes|cada mes)', re.I)
RX_ANTELACION = re.compile(
    r'(antes|post-|posterior|d[ií]a siguiente|[uú]ltimo d[ií]a|d[ií]a apertura|'
    r'v[ií]spera|al confirmar)', re.I)
#: DOM-R2-11 — «Día 1»…«Día 5» y «Fin de semana» son días de la SEMANA del
#: cliente, no cadencias: si votasen «Cadencia» (RX_CADENCIA casa «día \d») la
#: cabecera de «Semanal Manager» dejaría de decir «Día».
RX_DIA_N = re.compile(r'(?i)^d[ií]a\s*\d+$')
RX_FINDE = re.compile(r'(?i)^fin(es)? de semana$')

#: DOM-R2-21 — los dos BONUS firmaban sin el nombre del kit (y uno con un guion
#: suelto al final). El pie se reescribe entero desde CTX en cualquier celda que
#: se parezca a un pie, no sólo en la que decía «© AI Chef Pro».
#: §7-bis.17 — «— Kit de Tareas: <kit> · ChefBusiness …» es el pie de sushi-bar
#: y asador (13 celdas cada uno). Sin esta alternativa, `cerrar` no lo
#: reconocía como pie y la única razón por la que no quedaban DOS pies en la
#: hoja era que `reescribir_instrucciones` descarta las líneas que empiezan por
#: «—». Ningún kit de la familia escribe «Kit de Tareas:» con dos puntos
#: (medido: 0 celdas en los 13), así que la alternativa no los alcanza.
RX_PIE = re.compile(
    r'^\s*(©\s*AI Chef Pro|—\s*Kit de Tareas Recurrentes|—\s*Kit de Tareas:)')

#: DOM-R2-22 — el corpus mezclaba «−20 °C» (menos tipográfico + espacio) con
#: «-18°C» y «>65°C». Se normaliza a U+2212 + espacio ANTES de la unidad. El
#: guion de un RANGO («0-4 °C») va precedido de dígito y no se toca.
RX_GRADOS = re.compile(r'(?<=\d)\s*°\s*C\b')
RX_MENOS = re.compile(r'(?<![\d\w])-(?=\s?\d+(?:[.,]\d+)?\s?°C)')
#: DOM-R2-22 (2ª vuelta) — en un RANGO que comparte unidad al final («pasar de
#: -18 a −12 °C», «cámara pescado crudo (-2 a 0 °C)») el PRIMER signo no está
#: pegado al «°C», así que RX_MENOS no lo veía y se quedaba con guion ASCII
#: junto a un hermano ya tipográfico. Este patrón mira a través del conector.
#: Sigue sin tocar el guion de rango de «0-4 °C»: va precedido de dígito.
RX_MENOS_RANGO = re.compile(
    r'(?<![\d\w])-(?=\s?\d+(?:[.,]\d+)?\s*(?:a|y|hasta|hacia)\s*'
    r'[−-]?\s?\d+(?:[.,]\d+)?\s?°C)')

#: T-02 (tanda 4) — el mismo TPV se ENCENDÍA dos veces en el mismo kit: en el
#: fichero de NEGOCIO («Encender TPV / POS / datáfono», 07:15) y en el de CAJA
#: («Encender TPV / POS», 07:30), bajo una línea impresa que promete que los
#: ficheros no se duplican. El hito de encender es del LOCAL; lo que hace el
#: responsable de caja es comprobarlo y abrir su turno. Regla de FAMILIA: se
#: aplica sobre el fichero de caja de cualquier kit, no en los módulos de
#: contenido. La hora NO se toca (es el escalonado de `precargar_caja`).
#: Sólo pica cuando la tarea entera es «encender el TPV»: la de pastelería
#: («Encender el TPV y el datáfono; comprobar el rollo…», que además vive en el
#: fichero de negocio) no se toca.
RX_TPV_CAJA = re.compile(
    r'(?i)^\s*encender\s+(?:el\s+)?tpv(?:\s*/\s*pos)?(?:\s*/\s*dat[áa]fono)?'
    r'\s*$')
TXT_TPV_CAJA = 'Comprobar que el TPV está encendido y abrir turno de caja'

#: T-08 (tanda 4) — «No se duplican: cada uno cubre un nivel.» prometía más de
#: lo que el motor garantiza: `anotar_duplicados` mide solapes reales del 25-40 %
#: que el umbral del 80 % deja pasar. En vez de bajar el umbral (y borrar
#: contenido de un hermano a ciegas) se dice la verdad, con el mismo literal en
#: los 11 ficheros de todos los kits.
FRASE_NIVELES = (
    'Cada fichero cubre un nivel: el de negocio marca el HITO (encender, '
    'abrir, cerrar), el de áreas detalla CÓMO se hace en cada zona y el de '
    'caja lleva el DINERO. Si una tarea aparece en dos, es a propósito: una '
    'es el hito y la otra el detalle.')
#: m1 (tanda 5) — la frase de arriba enumera TRES niveles, y en los cinco kits
#: del molde P4 el kit sólo tiene dos ficheros en alcance: no hay fichero de
#: ÁREAS. La misma viñeta abría «Orden de uso: local → caja» (dos pasos) y
#: seguía «…el de áreas detalla CÓMO se hace en cada zona…» (tres ficheros): se
#: contradecía a sí misma en 10 celdas (catering 08!B38 y 09!B49, chocolatería
#: 08!B37 y 09!B48, heladería 08!B37 y 09!B48, hotel 18!B37 y 19!B48,
#: restaurante-creativo 10!B37 y 11!B48). Es la misma clase de defecto que T-01:
#: una rama que no comprueba el caso que la contradice. Se condiciona a
#: `CTX['f_areas']`, que es lo mismo que mira `_bloque_conecta` para decidir si
#: emite la línea de áreas y si mete «áreas» en el orden de uso.
FRASE_NIVELES_SIN_AREAS = (
    'Cada fichero cubre un nivel: el de negocio marca el HITO (encender, '
    'abrir, cerrar) y el de caja lleva el DINERO. Si una tarea aparece en los '
    'dos, es a propósito: una es el hito y la otra el detalle.')
#: CB-E7 (sub-familia ChefBusiness) — la tercera variante: kits SIN fichero de
#: negocio y SIN fichero de caja (sushi-bar, asador, chef-privado). Las dos
#: frases de arriba prometen un nivel de «negocio» y otro de «dinero» que estos
#: kits no entregan: no existen los 08/09 de la familia y ningún fichero suyo
#: tiene firma de recuento, registro mensual, liquidación ni registro de
#: eventos (medido el 2026-08-29 sobre `dl/`). Es la misma clase de defecto que
#: m1: una frase que enumera niveles que el kit no tiene.
FRASE_NIVELES_SOLO_AREAS = (
    'Este kit no trae un fichero aparte de negocio ni de caja: el de áreas es '
    'el que detalla CÓMO se hace el día en cada zona, y es también el marco '
    'del día. Si una tarea aparece en dos ficheros, es a propósito: uno la '
    'enuncia y el otro la detalla.')


#: m6 (motor 2.5) — MODELO DE CAJA POR EVENTOS. Una empresa de catering no
#: tiene mostrador: factura por EVENTO y cobra mayoritariamente por
#: transferencia (anticipo del 30-50 % + saldo tras el evento). Su fichero del
#: dinero no es un arqueo de cajón sino una liquidación por evento, así que las
#: frases de la familia que dicen «el de caja lleva el DINERO» describirían un
#: fichero que ya no existe. `CTX['modelo_caja']` vale 'mostrador' (los 10 kits
#: con arqueo) o 'eventos' (catering), y lo decide `contexto()` por CABECERA.
FRASE_NIVELES_EVENTOS = (
    'Cada fichero cubre un nivel: el de negocio marca el HITO (abrir, montar, '
    'cerrar) y el de cobros lleva el DINERO de cada evento (anticipo, factura '
    'y saldo). Si una tarea aparece en los dos, es a propósito: una es el hito '
    'y la otra el detalle.')
FRASE_NIVELES_EVENTOS_AREAS = (
    'Cada fichero cubre un nivel: el de negocio marca el HITO (abrir, montar, '
    'cerrar), el de áreas detalla CÓMO se hace en cada zona y el de cobros '
    'lleva el DINERO de cada evento (anticipo, factura y saldo). Si una tarea '
    'aparece en dos, es a propósito: una es el hito y la otra el detalle.')


def es_modelo_eventos():
    """m6 — ¿el fichero del dinero de este kit es de EVENTOS, no de mostrador?"""
    return CTX.get('modelo_caja') == 'eventos'


def frase_niveles():
    """m1 — la frase de T-08 con los niveles que este kit tiene DE VERDAD.

    m6 — y con el vocabulario del modelo de caja que tiene: en catering no hay
    «caja», hay cobros por evento.

    CB-E7 — y sin prometer un nivel de «negocio» o de «dinero» en los kits que
    no tienen ninguno de los dos. Hoy esta rama no la alcanza nadie desde
    `_bloque_conecta` (la frase sólo se imprime con `len(orden) > 1`, y sin
    negocio ni caja el `orden` tiene como mucho un paso, «áreas»); queda como
    la respuesta correcta para cualquier otro consumidor —el módulo de
    contenido de un kit, un bloque nuevo de Instrucciones— y para que la
    función no pueda devolver una frase falsa si mañana se la llama desde otro
    sitio. Está declarado como hallazgo en el informe de T0.
    """
    if CTX.get('sin_caja'):
        return FRASE_NIVELES_SOLO_AREAS
    if es_modelo_eventos():
        return (FRASE_NIVELES_EVENTOS_AREAS if CTX.get('f_areas')
                else FRASE_NIVELES_EVENTOS)
    return FRASE_NIVELES if CTX.get('f_areas') else FRASE_NIVELES_SIN_AREAS

#: R3-f — solape MEDIDO de cada banda contra el marco (08/09), se anote o no.
#: El umbral del ≥80 % no lo alcanza ninguna banda de los kits auditados (el
#: máximo medido es 40 %), así que sin esta lista la regla parecería aplicada
#: cuando en realidad no toca nada. `main.py` la vuelca en el informe.
SOLAPES = []

#: Registro de TODA fórmula escrita por el motor: `main.py` verifica una por una
#: que quedó con valor cacheado tras `inject_cache.py`.
REGISTRO = []

#: Contexto del KIT (lo rellena `contexto()` desde main.py antes de procesar):
#: nombre comercial, pie de hoja, sufijo de metadata, hora ancla de apertura,
#: literal de cierre y qué fichero hace de negocio / caja / áreas.
CTX = {}


def reg(ws, coord, formula):
    REGISTRO.append((ws.title, coord, formula))


# ==========================================================================
# Utilidades genéricas (portadas de kit-pasteleria-v2_0-postprocess.py)
# ==========================================================================
RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _nuevo_indice(i, idx, n):
    """Índice de fila/columna tras insertar (n>0) o borrar (n<0) en `idx`."""
    if n >= 0:
        return i + n if i >= idx else i
    if i >= idx - n:                 # idx + |n|
        return i + n
    if i >= idx:                     # dentro del hueco borrado
        return idx
    return i


def _traducir_formula(valor, idx, eje, n=1):
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        if eje == 'col':
            col = get_column_letter(
                _nuevo_indice(column_index_from_string(col), idx, n))
        else:
            fila = str(_nuevo_indice(int(fila), idx, n))
        return f'{d1}{col}{d2}{fila}'

    return RX_REF.sub(_sub, valor)


def _rangos_dv(ws):
    return [({k: getattr(dv, k, None) for k in CAMPOS_DV},
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None, n=1):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**{k: v for k, v in attrs.items() if v is not None})
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje, n) if idx else r)


def _desplazar_rango(ref, idx, eje, n=1):
    fuera = []
    for p in ref.split(':'):
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        if eje == 'col':
            col = get_column_letter(
                _nuevo_indice(column_index_from_string(col), idx, n))
        else:
            fila = str(_nuevo_indice(int(fila), idx, n))
        fuera.append(f'{d1}{col}{d2}{fila}')
    return ':'.join(fuera)


def insertar_columna(ws, idx):
    """Inserta una columna en `idx` moviendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos = {k: v.width for k, v in ws.column_dimensions.items() if v.width}

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra in sorted(anchos, key=column_index_from_string, reverse=True):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = anchos[letra]


def insertar_filas(ws, idx, n=1):
    """Equivalente por filas de `insertar_columna` (n filas de golpe)."""
    if n <= 0:
        return
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = {k: v.height for k, v in ws.row_dimensions.items() if v.height}

    for fila in range(max_r, idx - 1, -1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila + n, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila', n)
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'fila', n))
    _restaurar_dv(ws, dvs, idx, 'fila', n)
    _realturar(ws, alturas, idx, n)


def eliminar_filas(ws, idx, n=1):
    """Simétrica de `insertar_filas`: borra n filas desde `idx`.

    Hace falta para §2.5 (los bloques de 01!'Cierre Sala' que duplican 08 y 09
    se reducen a una línea). `ws.delete_rows` de openpyxl NO mueve merges, DV ni
    referencias de fórmula, que es justo lo que hay debajo del bloque.
    """
    if n <= 0:
        return
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = {k: v.height for k, v in ws.row_dimensions.items() if v.height}

    for fila in range(idx + n, max_r + 1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila - n, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila', -n)
            dst._style = copy.copy(src._style)
    # TEC-R2-15 — las n filas del final se quedan sin contenido pero CON el
    # estilo que arrastraban (relleno y bordes): la hoja seguía declarando
    # `dimensions` diez filas por debajo de su último dato y cualquier script
    # que recorra `ws.max_row` se despistaba. Se vacían valor Y estilo.
    for fila in range(max_r - n + 1, max_r + 1):
        for col in range(1, max_c + 1):
            cel = ws.cell(row=fila, column=col)
            cel.value = None
            cel.style = 'Normal'

    for m in merges:
        mi = [int(RX_REF.fullmatch(p).group(4)) for p in m.split(':')
              if RX_REF.fullmatch(p)]
        if mi and all(idx <= f < idx + n for f in mi):
            continue                       # combinación dentro del hueco
        ws.merge_cells(_desplazar_rango(m, idx, 'fila', -n))
    _restaurar_dv(ws, dvs, idx, 'fila', -n)
    _realturar(ws, alturas, idx, -n)


def _realturar(ws, alturas, idx, n):
    for f in list(alturas):
        ws.row_dimensions[f].height = None
    for f, alto in alturas.items():
        if n < 0 and idx <= f < idx - n:
            continue
        ws.row_dimensions[_nuevo_indice(f, idx, n)].height = alto


def print_setup(ws, header_row=None, landscape=True):
    ws.page_setup.paperSize = 9                      # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = f'{header_row}:{header_row}'
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def ancho_util(ws):
    """Suma de anchos de las columnas con contenido (unidades de Excel)."""
    total = 0.0
    for c in range(1, ws.max_column + 1):
        total += ws.column_dimensions[get_column_letter(c)].width or 8.43
    return total


def area_impresion(ws):
    """`print_area` hasta la última celda con contenido (COM-25/TEC-24)."""
    max_r = max_c = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                max_r, max_c = max(max_r, c.row), max(max_c, c.column)
    if not max_r:
        return
    ws.print_area = f'A1:{get_column_letter(max_c)}{max_r}'


def _verde(cel, fmt=None):
    cel.fill = PatternFill('solid', fgColor=VERDE)
    if fmt:
        cel.number_format = fmt


def _relleno(cel, color):
    cel.fill = PatternFill('solid', fgColor=color)


def _buscar(ws, texto, col=None):
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if c.value == texto:
                return c.row
    return None


def _buscar_prefijo(ws, prefijo, col=1):
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if isinstance(c.value, str) and c.value.startswith(prefijo):
                return c.row
    return None


def _merge(ws, ref):
    if ':' not in ref:
        return
    vivos = [str(r) for r in ws.merged_cells.ranges]
    if ref not in vivos:
        ws.merge_cells(ref)


def _desmerge_fila(ws, fila):
    for m in [str(r) for r in ws.merged_cells.ranges]:
        if all(int(RX_REF.fullmatch(p).group(4)) == fila
               for p in m.split(':') if RX_REF.fullmatch(p)):
            ws.unmerge_cells(m)


# ==========================================================================
# DETECCIÓN POR CABECERA — lo único que decide qué se toca
# ==========================================================================
#: Se detecta por PREFIJO, no por igualdad: la v2.0 renombra la columna del
#: efectivo («Ventas Efectivo» → «Efectivo Contado», DOM-R2-01) y con igualdad
#: exacta la 2.ª pasada dejaría de reconocer la hoja — el registro mensual se
#: publicaría sin descuadre, sin ámbar y sin protección.
CAB_REGISTRO = ('Fecha', 'Fondo Apertura', 'Total Facturado')
CAB_RECUENTO = ('Denominación', 'Cantidad')
CAB_CALENDARIO = ('Mes', 'Fecha / Evento', 'Tareas Clave', 'Antelación')
#: Cabecera EXACTA de las hojas de caja antes de la v2.0 (las 26 hojas de los
#: 13 kits de la familia). `caja_columnas` reordena por posición y sólo actúa
#: si la encuentra tal cual.
CAB_CAJA = ('#', 'Tarea', 'Responsable', '✓ Completada', 'Hora', 'Notas')

# --------------------------------------------------------------------------
# m6 — vocabulario del MODELO DE CAJA POR EVENTOS (09 de catering)
# --------------------------------------------------------------------------
#: Estas etiquetas son el CONTRATO entre tres piezas que tienen que decir
#: exactamente lo mismo o el fichero se rompe en silencio:
#:   · `construir_09_catering.py`, que las escribe;
#:   · `fila_liquidacion` / `fila_registro_eventos`, que detectan el papel del
#:     fichero por CABECERA (regla de oro: nunca por nombre de fichero);
#:   · `demo_liquidacion` de `main.py`, que localiza las celdas por su rótulo
#:     para demostrar §6 con pycel.
#: Por eso viven aquí y no en el constructor: si el constructor cambiara un
#: rótulo por su cuenta, el motor dejaría de reconocer el fichero, `papel` sería
#: None, `f_caja` quedaría vacío y las Instrucciones de los 11 ficheros del kit
#: se publicarían sin la línea del dinero, sin que ningún gate lo cantase.
#: Espacios NORMALES (U+0020) en «10 %» y «21 %», como el resto de la familia
#: («0-4 °C», «0,05 €»): el espacio fino U+202F es de los .md ensamblados.
ETIQ_EV_EVENTO = 'Evento / Cliente'
ETIQ_EV_FECHA = 'Fecha del evento'
ETIQ_EV_PAX_CONTR = 'Comensales contratados'
ETIQ_EV_PAX_REAL = 'Comensales reales'
ETIQ_EV_PRESUPUESTO = 'Presupuesto aceptado (base, sin IVA)'
ETIQ_EV_EXTRAS = 'Extras (base, sin IVA)'
ETIQ_EV_BASE10 = 'Base imponible al 10 %'
ETIQ_EV_BASE21 = 'Base imponible al 21 %'
ETIQ_EV_AVISO_BASE = 'Comprobación: las dos bases deben sumar presupuesto + extras'
ETIQ_EV_IVA10 = 'IVA 10 %'
ETIQ_EV_IVA21 = 'IVA 21 %'
ETIQ_EV_TOTAL = 'TOTAL FACTURA'
ETIQ_EV_ANTICIPO = 'Anticipo cobrado (−)'
ETIQ_EV_SALDO = 'Saldo tras anticipo'
ETIQ_EV_COBRADO = 'Cobrado tras el evento'
ETIQ_EV_PENDIENTE = 'PENDIENTE DE COBRO'
ETIQ_EV_VENCIMIENTO = 'Fecha de vencimiento del saldo'
ETIQ_EV_ESTADO = 'ESTADO'
#: Los tres estados que emite la fórmula de ESTADO, en la hoja de liquidación y
#: en el registro. Son literales de fórmula: cambiarlos aquí y no allí dejaría
#: al gate comparando contra un texto que el Excel no escribe nunca.
EV_COBRADO = 'Cobrado'
EV_PENDIENTE = 'Pendiente'
EV_VENCIDO = 'VENCIDO'
#: Tolerancia del «saldo cero»: un IVA del 10/21 % sobre bases con decimales no
#: da 0 exacto casi nunca. 0,01 € = un céntimo, que es lo que factura.
EV_TOLERANCIA = 0.01
#: Cabecera del «Registro de Eventos» (una fila por evento). Se detecta por
#: PREFIJO, como `CAB_REGISTRO`: «Base (presupuesto + extras)» y «Total factura»
#: llevan cola. No comparte firma con el registro mensual de caja, que exige
#: «Fondo Apertura» — un concepto que en catering no existe.
CAB_EVENTOS = ('Fecha', 'Evento / Cliente', 'Anticipo', 'Cobrado', 'Pendiente',
               'Estado')
#: Rojo de aviso del modelo por eventos. La familia sólo tenía ÁMBAR
#: (`AMBAR`, el descuadre del arqueo), y un saldo VENCIDO es otra cosa: el
#: ámbar dice «queda por cobrar» y el rojo «ya pasó el plazo, llama hoy».
#: Pintar los dos igual borraría la única distinción que el fichero hace.
ROJO = 'F8D7DA'
ROJO_TXT = '721C24'


def cf_eventos(ws, tipo):
    """(Re)aplica el formato condicional de las dos hojas del modelo eventos.

    Vive AQUÍ y no en `construir_09_catering.py` por un defecto medido el
    2026-08-23: `aplicar` VACÍA `ws.conditional_formatting` de todas las hojas
    reconocidas para que el pipeline sea idempotente, y sólo lo reconstruye
    para los checklists (§2) y para el arqueo de mostrador. El ámbar del
    PENDIENTE y el rojo del VENCIDO existían en el fichero recién construido y
    DESAPARECÍAN en la primera pasada del motor, con el dry-run en verde y sin
    que ningún gate lo cantase — el mismo patrón que la caché del frontmatter
    del blog: el cambio parece aplicado porque el resto de la hoja sí lo está.

    Las celdas se localizan por RÓTULO, nunca por coordenada, para que el
    constructor pueda mover una fila sin romper esto. Devuelve el nº de reglas.
    """
    reglas = 0
    ambar = PatternFill('solid', start_color=AMBAR, end_color=AMBAR)
    rojo = PatternFill('solid', start_color=ROJO, end_color=ROJO)
    f_rojo = Font(bold=True, color='00' + ROJO_TXT)
    if tipo == 'liquidacion':
        pend = _buscar(ws, ETIQ_EV_PENDIENTE, col=1) or _buscar(
            ws, ETIQ_EV_PENDIENTE, col=2)
        estado = _buscar(ws, ETIQ_EV_ESTADO, col=1) or _buscar(
            ws, ETIQ_EV_ESTADO, col=2)
        if pend:
            ws.conditional_formatting.add(
                f'C{pend}', CellIsRule(operator='greaterThan',
                                       formula=[str(EV_TOLERANCIA)],
                                       fill=ambar))
            reglas += 1
        if estado:
            ws.conditional_formatting.add(
                f'C{estado}', CellIsRule(operator='equal',
                                         formula=[f'"{EV_VENCIDO}"'],
                                         fill=rojo, font=f_rojo))
            reglas += 1
        return reglas
    if tipo != 'registro_eventos':
        return 0
    hr = fila_registro_eventos(ws)
    if not hr:
        return 0
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hr, column=c).value
        if isinstance(v, str) and v.strip():
            cols[v.strip()] = c
    def _col(prefijo):
        for rot, c in cols.items():
            if rot == prefijo or rot.startswith(prefijo):
                return c
        return None
    c_pend, c_estado = _col('Pendiente'), _col('Estado')
    # El cuerpo acaba en la fila anterior a TOTALES; si algún día no la
    # hubiera, en la última fila con número de fila de datos.
    fin = _buscar(ws, 'TOTALES', col=1)
    ultima = (fin - 1) if fin else ws.max_row
    if ultima <= hr:
        return 0
    if c_pend:
        letra = get_column_letter(c_pend)
        ws.conditional_formatting.add(
            f'{letra}{hr + 1}:{letra}{ultima}',
            CellIsRule(operator='greaterThan', formula=[str(EV_TOLERANCIA)],
                       fill=ambar))
        reglas += 1
    if c_estado:
        # La fila ENTERA en rojo: en una tabla de 25 líneas, un color en una
        # sola celda se pierde justo cuando más falta hace verlo.
        letra = get_column_letter(c_estado)
        ws.conditional_formatting.add(
            f'A{hr + 1}:{get_column_letter(ws.max_column)}{ultima}',
            FormulaRule(formula=[f'${letra}{hr + 1}="{EV_VENCIDO}"'],
                        fill=rojo))
        reglas += 1
    return reglas


def cabecera_checklist(ws):
    """(fila, {texto: columna}) de la cabecera de un checklist de la familia."""
    for r in range(3, 8):
        if ws.cell(row=r, column=2).value != 'Tarea':
            continue
        if ws.cell(row=r, column=1).value not in ('Nº', '#'):
            continue
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                cols[v.strip()] = c
        if CAB_MARCA in cols:
            return r, cols
    return None, None


#: Rótulo de la fila de totales. NO es literal: el kit base escribe «Tareas
#: completadas:» en A o en B y el de cafetería «Completadas:» en C. Buscar el
#: literal del kit base dejaba a los hermanos sin contador (y sin las 5 filas
#: libres, que se miden contra esa fila).
RX_CONTADOR = re.compile(r'(?i)^\s*(tareas\s+)?completadas\s*:?\s*$')
ETIQ_CONTADOR = 'Tareas completadas:'


def geometria(ws):
    """Geometría de una hoja de checklist, o None si no lo es."""
    hr, cols = cabecera_checklist(ws)
    if hr is None:
        return None
    contador = None
    for r in range(hr + 1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 4) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and RX_CONTADOR.match(v):
                contador = r
                break
        if contador:
            break
    ultima = None
    for r in range(hr + 1, contador or ws.max_row + 1):
        if isinstance(ws.cell(row=r, column=1).value, int):
            ultima = r
    if ultima is None:
        return None
    return {'hr': hr, 'cols': cols, 'marca': cols[CAB_MARCA],
            'contador': contador, 'ultima': ultima}


def fila_registro_mensual(ws):
    for r in range(3, 8):
        fila = [v.strip() for v in
                (ws.cell(row=r, column=c).value
                 for c in range(1, min(ws.max_column, 12) + 1))
                if isinstance(v, str) and v.strip()]
        if all(any(t == v or v.startswith(t) for v in fila)
               for t in CAB_REGISTRO):
            return r
    return None


def fila_recuento(ws):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == CAB_RECUENTO[0] and \
           ws.cell(row=r, column=2).value == CAB_RECUENTO[1]:
            return r
    return None


def fila_registro_eventos(ws):
    """Fila de cabecera del «Registro de Eventos» (m6), o None.

    Misma técnica que `fila_registro_mensual` —prefijo, no igualdad— y por el
    mismo motivo: los rótulos de esa hoja llevan cola («Base (presupuesto +
    extras)», «Total factura») y con igualdad exacta cualquier retoque futuro
    dejaría la hoja fuera de alcance sin un solo aviso.
    """
    for r in range(3, 9):
        fila = [v.strip() for v in
                (ws.cell(row=r, column=c).value
                 for c in range(1, min(ws.max_column, 14) + 1))
                if isinstance(v, str) and v.strip()]
        if all(any(t == v or v.startswith(t) for v in fila)
               for t in CAB_EVENTOS):
            return r
    return None


def fila_liquidacion(ws):
    """Fila del «PENDIENTE DE COBRO» de la «Liquidación del Evento», o None.

    No es una tabla con cabecera sino un FORMULARIO (etiqueta a la izquierda,
    importe en una columna verde), así que lo que la identifica es la pareja de
    rótulos que sólo aparece junta ahí: «TOTAL FACTURA» y «PENDIENTE DE COBRO».
    Se buscan en las dos primeras columnas porque la etiqueta va combinada A:B,
    igual que el Resumen de Cierre del modelo de mostrador (TEC-03).
    """
    total = pendiente = None
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            v = v.strip()
            if v == ETIQ_EV_TOTAL:
                total = r
            elif v == ETIQ_EV_PENDIENTE:
                pendiente = r
    return pendiente if (total and pendiente) else None


def fila_calendario(ws):
    """Fila de cabecera de un calendario ▸, o None.

    Hay DOS moldes en la familia y el segundo estaba fuera del motor sin que
    nadie lo hubiera decidido: el BONUS-02 de bar es «# | Fecha | Evento |
    Preparación Especial | Antelación | Notas» y quedaba como el único fichero
    del kit sin bio, sin A4 y con el pie viejo (10/11 tocados frente a los 11
    del resto). Lo que define la familia es la pareja «Antelación» + una
    columna de EVENTO; los calendarios de catering («Fecha / Período | Tipo»)
    y de hotel («Fecha / Período | Impacto») NO la tienen y siguen fuera, que
    es justo lo que pide el alcance «sólo 08/09» de esos cinco kits.
    """
    for r in range(1, 9):
        txt = [v.strip() for v in
               (ws.cell(row=r, column=c).value
                for c in range(1, min(ws.max_column, 10) + 1))
               if isinstance(v, str) and v.strip()]
        if 'Antelación' not in txt:
            continue
        if 'Fecha / Evento' in txt or 'Evento' in txt:
            return r
    return None


#: CB-E3 — vocabulario del molde REGISTRO. Un registro APPCC no es un
#: checklist: no tiene columna «Tarea» ni marca de completada, tiene una fila
#: por LOTE o por EQUIPO con fechas, temperaturas, lotes y firmas.
RX_REG_CAMPO = re.compile(
    r'(?i)^(fecha|temp|caducidad|firma|verif|lote|especie|proveedor|equipo|'
    r'inicio|fin)')
RX_REG_FECHA = re.compile(
    r'(?i)^(fecha|caducidad|inicio|fin|entrada|salida|compra)')
RX_REG_TEMP = re.compile(r'(?i)(temp|°\s*c)')
RX_REG_VERIF = re.compile(r'(?i)^verif')
DIAS_SEMANA = ('lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado',
               'domingo')


def fila_registro_appcc(ws):
    """CB-E3 — (fila, {rótulo: columna}) de un REGISTRO APPCC, o (None, None).

    Firma estructural, medida sobre las tres hojas de
    `kit-tareas-sushi-bar/03-seguridad-anisakis-appcc.xlsx` y sobre
    `kit-tareas-marisqueria/03-trazabilidad-appcc-marisco.xlsx`:
    «#»/«Nº» en la columna A, cinco rótulos o más, **ninguna columna «Tarea»**
    (eso sería el molde ▸ o el P4) y al menos un campo de registro — o siete
    días de la semana, que es como `'Temperaturas Diario'` monta su rejilla.

    Se excluye a propósito lo que lleva «Antelación»: son los BONUS-02 de
    catering y de hotel («# | Fecha / Período | Impacto | Preparación F&B |
    Antelación | Notas»), calendarios de dos kits LIVE que casarían con todo lo
    demás. Sin esa exclusión, el gate de regresión de hotel se habría puesto
    rojo en la primera pasada.
    """
    for r in range(2, 9):
        if ws.cell(row=r, column=1).value not in ('#', 'Nº'):
            continue
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                cols[v.strip()] = c
        if len(cols) < 5:
            continue
        rot = [_sin_tildes(k) for k in cols]
        if 'tarea' in rot or 'antelacion' in rot:
            continue
        dias = sum(1 for d in DIAS_SEMANA if d in rot)
        if not (any(RX_REG_CAMPO.match(k) for k in cols) or dias >= 5):
            continue
        # tiene que haber cuerpo: al menos dos filas numeradas debajo
        n = sum(1 for rr in range(r + 1, ws.max_row + 1)
                if isinstance(ws.cell(row=rr, column=1).value, int))
        if n < 2:
            continue
        return r, cols
    return None, None


#: Rangos de temperatura escritos DENTRO del rótulo de la fila. Se leen del
#: propio texto —«(−2 a 0 °C)», «(2-4 °C)», «(−20 °C o menos)»— porque el
#: límite es del EQUIPO, no del kit: cada fila declara el suyo. La coherencia
#: entre ficheros (gate `limite_unico`, §7-bis.23) la arregla el módulo de
#: contenido ANTES; aquí sólo se pinta lo que la fila ya dice.
RX_RANGO = re.compile(
    r'(?<![\d])([−\-]?\d+(?:[.,]\d+)?)\s*(?:°\s*C)?\s*(?:a|y|hasta|/)\s*'
    r'([−\-]?\d+(?:[.,]\d+)?)\s*°\s*C')
RX_RANGO_GUION = re.compile(
    r'(?<![\d\-−])(\d+(?:[.,]\d+)?)\s*-\s*(\d+(?:[.,]\d+)?)\s*°\s*C')
RX_TOPE = re.compile(
    r'([−\-]?\d+(?:[.,]\d+)?)\s*°\s*C\s*(o\s+menos|o\s+inferior|m[aá]ximo)')
RX_SUELO = re.compile(
    r'([−\-]?\d+(?:[.,]\d+)?)\s*°\s*C\s*(o\s+m[aá]s|o\s+superior|m[ií]nimo)')


def _num(v):
    return float(v.replace('−', '-').replace(',', '.'))


def rango_de_texto(v):
    """(min, max) de temperatura declarados en el texto, o None. min/max pueden
    ser None cuando el texto sólo fija un tope o un suelo."""
    if not isinstance(v, str):
        return None
    m = RX_RANGO.search(v) or RX_RANGO_GUION.search(v)
    if m:
        a, b = _num(m.group(1)), _num(m.group(2))
        return (min(a, b), max(a, b))
    m = RX_TOPE.search(v)
    if m:
        return (None, _num(m.group(1)))
    m = RX_SUELO.search(v)
    if m:
        return (_num(m.group(1)), None)
    return None


def registro_appcc(ws, cambios):
    """CB-E3 — formatos, desplegable, CF de rango y alturas de un REGISTRO.

    Cierra TEC-07 (fechas sin formato), TEC-08 (columna de verificación sin
    desplegable) y TEC-19 (`row_dimensions` vacío: la cabecera de nueve
    columnas se imprimía en una línea recortada). Devuelve el cuerpo
    (primera, última) para que `proteger` lo desbloquee.
    """
    hr, cols = fila_registro_appcc(ws)
    if hr is None:
        return None
    ncol = max(cols.values())
    ultima = hr
    for r in range(hr + 1, ws.max_row + 1):
        if isinstance(ws.cell(row=r, column=1).value, int):
            ultima = r
    if ultima <= hr:
        return None

    rot = {c: k for k, c in cols.items()}
    dias = sum(1 for d in DIAS_SEMANA if d in [_sin_tildes(k) for k in cols])
    col_etiqueta = 2 if 2 in rot else None
    # La rejilla semanal («# | Equipo / Cámara | Lunes … Domingo») no dice
    # «Temp» en ninguna cabecera: lo que lleva la temperatura son los días.
    temp_por_dias = dias >= 5 and col_etiqueta is not None

    fechas = tempes = verif = 0
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if dv.type != 'list']
    dv_marca = None
    for c in range(1, ncol + 1):
        etq = rot.get(c, '')
        es_temp = bool(RX_REG_TEMP.search(etq)) or (
            temp_por_dias and _sin_tildes(etq) in DIAS_SEMANA)
        for r in range(hr + 1, ultima + 1):
            cel = ws.cell(row=r, column=c)
            if c == 1:
                continue
            _verde(cel)
            if RX_REG_FECHA.match(etq):
                cel.number_format = 'DD/MM/YYYY'
            elif es_temp:
                cel.number_format = '0.0 "°C"'
        if RX_REG_FECHA.match(etq):
            fechas += 1
        elif es_temp:
            tempes += 1
        if RX_REG_VERIF.match(etq):
            if dv_marca is None:
                dv_marca = DataValidation(
                    type='list', formula1=DV_LISTA, allow_blank=True,
                    showErrorMessage=True, errorStyle='stop',
                    errorTitle=DV_ERROR_TIT, error=DV_ERROR)
                ws.add_data_validation(dv_marca)
            letra = get_column_letter(c)
            dv_marca.add(f'{letra}{hr + 1}:{letra}{ultima}')
            verif += 1

    # --- CF de fuera de rango, fila a fila y con el límite que la fila dice --
    ws.conditional_formatting = ConditionalFormattingList()
    rojo = PatternFill('solid', start_color=ROJO, end_color=ROJO)
    reglas = sin_rango = 0
    if temp_por_dias:
        cdias = sorted(c for c, k in rot.items()
                       if _sin_tildes(k) in DIAS_SEMANA)
        if cdias:
            ini, fin = get_column_letter(cdias[0]), get_column_letter(cdias[-1])
            for r in range(hr + 1, ultima + 1):
                rango = rango_de_texto(ws.cell(row=r, column=col_etiqueta)
                                       .value)
                if not rango:
                    sin_rango += 1
                    continue
                lo, hi = rango
                ref = f'{ini}{r}:{fin}{r}'
                if lo is not None and hi is not None:
                    ws.conditional_formatting.add(
                        ref, CellIsRule(operator='notBetween',
                                        formula=[str(lo), str(hi)], fill=rojo))
                elif hi is not None:
                    ws.conditional_formatting.add(
                        ref, CellIsRule(operator='greaterThan',
                                        formula=[str(hi)], fill=rojo))
                else:
                    ws.conditional_formatting.add(
                        ref, CellIsRule(operator='lessThan',
                                        formula=[str(lo)], fill=rojo))
                reglas += 1

    # --- TEC-19: la cabecera con wrap y alto explícito ---------------------
    for c in range(1, ncol + 1):
        cel = ws.cell(row=hr, column=c)
        cel.alignment = Alignment(wrap_text=True, vertical='center',
                                  horizontal=cel.alignment.horizontal
                                  or 'center')
    largo = max((len(str(rot.get(c, ''))) for c in range(1, ncol + 1)),
                default=0)
    ancho = min(ws.column_dimensions[get_column_letter(c)].width or 8.43
                for c in range(2, ncol + 1))
    ws.row_dimensions[hr].height = max(30, 15 * (int(largo / max(ancho, 6)) + 1))

    # --- CB-E9 (b): la columna del EQUIPO, ancha de verdad -----------------
    if col_etiqueta:
        letra = get_column_letter(col_etiqueta)
        necesario = max((len(str(ws.cell(row=r, column=col_etiqueta).value))
                         for r in range(hr + 1, ultima + 1)
                         if isinstance(ws.cell(row=r, column=col_etiqueta)
                                       .value, str)), default=0)
        actual = ws.column_dimensions[letra].width or 8.43
        tope = min(ANCHO_TAREA_CB, actual + (ANCHO_A4 - ancho_util(ws)))
        if necesario > actual and tope > actual:
            ws.column_dimensions[letra].width = round(min(necesario + 2, tope))
            cambios.append(
                f'CB-E9 «{ws.title}»: columna {letra} de {actual:g} a '
                f'{ws.column_dimensions[letra].width:g} (el rótulo más largo '
                f'mide {necesario} caracteres)')

    cambios.append(
        f'CB-E3 «{ws.title}»: registro APPCC normalizado ({fechas} columnas de '
        f'fecha DD/MM/YYYY, {tempes} de temperatura 0,0 °C, {verif} con '
        f'desplegable, {reglas} reglas de fuera-de-rango'
        + (f', {sin_rango} filas sin límite legible en su rótulo' if sin_rango
           else '') + ')')
    return (hr + 1, ultima)


def ncol_cabecera(ws, hr):
    """Última columna con rótulo en la fila de cabecera `hr`."""
    n = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hr, column=c).value
        if isinstance(v, str) and v.strip():
            n = c
    return n or ws.max_column


def es_briefing(ws):
    b1, b2 = ws.cell(row=1, column=2).value, ws.cell(row=2, column=2).value
    return (isinstance(b1, str) and 'Briefing' in b1
            and isinstance(b2, str) and b2.startswith('Fecha:'))


def hojas_reconocidas(wb):
    """{título: tipo} de las hojas que este motor sabe tratar."""
    fuera = {}
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        if geometria(ws):
            fuera[ws.title] = 'checklist'
        elif fila_registro_mensual(ws):
            fuera[ws.title] = 'registro'
        # m6 — las dos hojas del dinero del modelo POR EVENTOS. Van DESPUÉS del
        # registro mensual (que exige «Fondo Apertura», ausente aquí) y ANTES
        # del calendario. Sin reconocerlas, `cerrar` no les daría A4, ni
        # print_area, ni protección, ni el pie del kit: se publicarían como las
        # dos únicas hojas sueltas de un producto v2.0.
        elif fila_registro_eventos(ws):
            fuera[ws.title] = 'registro_eventos'
        elif fila_liquidacion(ws):
            fuera[ws.title] = 'liquidacion'
        elif fila_calendario(ws):
            fuera[ws.title] = 'calendario'
        elif es_briefing(ws):
            fuera[ws.title] = 'briefing'
        # CB-E3 — el molde REGISTRO va el ÚLTIMO de la cadena a propósito:
        # así ninguna hoja que el motor ya sabía tratar puede cambiar de tipo
        # por su culpa, y la extensión sólo puede AÑADIR alcance.
        elif sub_cb() and fila_registro_appcc(ws)[0]:
            fuera[ws.title] = 'registro_appcc'
    return fuera


def en_alcance(wb):
    return bool(hojas_reconocidas(wb))


def es_fila_seccion(ws, fila):
    """Banda de sección = celda combinada que arranca en A y cubre ≥3 columnas."""
    for m in ws.merged_cells.ranges:
        if (m.min_row == fila == m.max_row and m.min_col == 1
                and m.max_col >= 3):
            return True
    return False


# ==========================================================================
# CONTEXTO del kit (se calcula una vez por carpeta, en main.py)
# ==========================================================================
def _mayoria(valores):
    """Valor más repetido, con desempate ALFABÉTICO.

    `max(set(x), key=x.count)` depende del orden de iteración de un `set` de
    cadenas y, con `PYTHONHASHSEED` aleatorio, dos ejecuciones pueden elegir
    candidatos distintos ante un empate: el gate de idempotencia lo vería como
    una diferencia intermitente imposible de reproducir.
    """
    if not valores:
        return ''
    return sorted(set(valores), key=lambda v: (-valores.count(v), v))[0]


class KitAmbiguo(RuntimeError):
    """Dos ficheros compiten por el mismo papel: se ABORTA, no se adivina."""


class MoldeDesconocido(RuntimeError):
    """CB-E6 — el producto entero es de un molde que el motor no conoce.

    Hermana de `KitAmbiguo` y por el mismo motivo: el motor no adivina. Hasta
    ahora un kit del molde PLANO (una hoja por fichero, cabecera en la fila 3,
    sin columna «Nº», sin contador y sin hoja «Instrucciones») atravesaba el
    pipeline entero **en verde**: `hojas_reconocidas` devolvía `{}` en los 11
    ficheros, `aplicar` caía en la rama P4, que tampoco reconocía nada,
    `cerrar` no llegaba a correr — y aun así `main.py` GUARDABA los 11 ficheros
    con la metadata nueva y el censo `--fail` daba 0 defectos. Medido el
    2026-08-29 en `kit-tareas-tapas-bar`: «en alcance 0/11», 0 fórmulas, 11
    cambios de sólo metadata. Un producto que sale «guardado sin novedad»
    porque el motor no ha entendido NADA de él es la peor salida posible: el
    informe dice verde y el entregable se queda en la v1.1 con el sello de la
    v2.0.
    """


def papel_del_fichero(wb):
    """(papel, detalle) de un fichero del kit, SÓLO por estructura.

    R3-a. Antes el papel salía de «¿cuántos checklists de apertura/cierre
    tiene?» y en bar y en dark-kitchen el 01 (Apertura Barra / Cierre Barra)
    tiene exactamente DOS, así que se llevaba el papel de NEGOCIO delante del
    08 real: DOM-06 se aplicaba al fichero equivocado, el gate lo daba por
    verde y en bar el nombre del kit crecía en cada pasada.

    Lo que distingue al fichero del LOCAL no es el número de hojas sino su
    molde: es el único con la columna «Notas» (8 columnas) en sus dos
    checklists de apertura/cierre. El de CAJA es el del recuento por
    denominaciones o el del registro mensual. El de ÁREAS, el que trae tres o
    más ciclos sin «Notas» (y, si el kit no tiene ninguno así, el de dos: en
    bar y dark-kitchen el detalle por área vive en un solo fichero de dos
    hojas).
    """
    recon = hojas_reconocidas(wb)
    if not recon:
        return None, {}
    ciclo = [t for t, k in recon.items() if k == 'checklist'
             and (t.lower().startswith('apertura')
                  or t.lower().startswith('cierre'))]
    con_notas = []
    for t in ciclo:
        g = geometria(wb[t])
        if g and 'Notas' in g['cols']:
            con_notas.append(t)
    detalle = {'ciclo': sorted(ciclo), 'con_notas': sorted(con_notas),
               'recuento': _tiene(wb, fila_recuento),
               'registro': _tiene(wb, fila_registro_mensual),
               # m6 — firma del modelo POR EVENTOS
               'liquidacion': _tiene(wb, fila_liquidacion),
               'registro_eventos': _tiene(wb, fila_registro_eventos)}
    # m6 — el fichero del dinero de catering: liquidación y registro POR
    # EVENTO. Se mira ANTES que el arqueo de mostrador, y no es un detalle de
    # estilo: el 09 de eventos lleva una sección OPCIONAL «Solo si hubo barra
    # con cobro en efectivo» con la misma tabla «Denominación | Cantidad» del
    # modelo de mostrador, así que `fila_recuento` lo encuentra y, con el orden
    # inverso, el fichero se llevaba el papel 'caja': `precargar_caja` le
    # habría escrito «Responsable de caja» y horas de reloj encima de los
    # responsables y de los D-15/D+7, `instrucciones_caja` le habría puesto un
    # manual de arqueo y §6 habría intentado demostrar un descuadre que ese
    # fichero no calcula. La firma de EVENTOS es específica (sólo aparece ahí);
    # la del recuento la comparten los dos modelos. Medido el 2026-08-23 con el
    # 09 de prueba: con el orden viejo, papel = 'caja'.
    if detalle['liquidacion'] or detalle['registro_eventos']:
        return 'cobros', detalle
    if detalle['recuento'] or detalle['registro']:
        return 'caja', detalle
    if len(ciclo) == 2 and len(con_notas) == 2:
        return 'negocio', detalle
    if len(ciclo) >= 3:
        return 'areas', detalle
    if len(ciclo) == 2:
        return 'areas2', detalle
    return None, detalle


def contexto(carpeta, ficheros, abrir, producto=None):
    """Rellena CTX leyendo los ficheros del kit. `abrir(fname)` → Workbook.

    Sólo lee. Recorre TODOS los ficheros (también los que el motor no toca: de
    ellos salen la hora ancla del kit de hotel, cuyos 01-17 son otro molde),
    pero `ficheros` en el contexto lista únicamente los que están en alcance.
    """
    # CB-E3/CB-E9 — el identificador del producto tiene que estar puesto ANTES
    # de leer un solo fichero: `hojas_reconocidas` consulta `sub_cb()` para
    # decidir si el molde REGISTRO está activo, y `sub_cb()` mira `CTX`. Hasta
    # esta tanda `main.py` lo asignaba DESPUÉS de `contexto()`, y el gate de
    # idempotencia lo cazó con 34 diferencias: en la 1.ª pasada el contexto de
    # sushi-bar clasificaba `03-seguridad-anisakis-appcc.xlsx` como
    # «checklist» (CTX aún vacío) y en la 2.ª como «registro» (CTX ya
    # heredado del proceso anterior), así que sus «Instrucciones» salían con
    # el bloque «Cómo personalizar» EQUIVOCADO la primera vez.
    if producto:
        CTX['producto'] = producto
    ctx = {'kit': '', 'pie': '', 'sufijo': '', 'hora_apertura': '07:00',
           'literal_cierre': 'Cierre', 'f_negocio': None, 'f_caja': None,
           'f_areas': None, 'ficheros': [], 'con_checklist': set(),
           # DOM-R2-17/COM-R2-13 — para que «Se conecta con» diga algo en los
           # 7 ficheros que no son ni el local, ni las áreas, ni la caja.
           'f_calendario': None, 'f_briefing': None, 'f_plantilla': None,
           'f_eventos': None, 'f_periodico': None,
           # T-04 — de aquí salen los paréntesis de «Se conecta con». Si el kit
           # no permite derivarlos, quedan vacíos y la frase va SIN paréntesis:
           # dark-kitchen no tiene terraza y su 01 sólo abre y cierra cocina.
           'areas_nombres': [], 'negocio_bandas': [],
           # m6 — 'mostrador' (arqueo de cajón: 10 kits) o 'eventos'
           # (liquidación por evento: catering). Lo decide la CABECERA del
           # fichero del dinero, no el nombre del producto.
           'modelo_caja': 'mostrador',
           # CB-E7 — kit SIN fichero de negocio y SIN fichero de caja (y sin
           # ninguna firma del dinero en ningún fichero, esté o no en alcance).
           # Lo rellena esta misma función, al final.
           'sin_caja': False,
           # CB-E4 — {fichero: True} si sus «Instrucciones» traen el bloque
           # «Cómo personalizar» que el motor sustituye, y {fichero: papel}
           # con el TIPO de entregable (checklist / registro / formulario /
           # calendario) con el que se redacta ese bloque.
           'personalizar': {}, 'papel_instrucciones': {}}
    horas, cierres, sufijos, kits = {}, {}, [], []
    #: §7-bis.17 / regla de marca — ¿este kit lleva la razón social de
    #: ChefBusiness? Se decide por DATO, no por lista de productos: los 13 kits
    #: de la familia no la mencionan NI UNA vez (medido sobre `dl/` el
    #: 2026-08-29), y de la sub-familia la llevan seis (13 celdas en sushi-bar
    #: y asador, 22 en marisquería, panadería, food-truck y tapas-bar) y
    #: chef-privado ninguna.
    marca_cb = False
    #: CB-E6/CB-E7 — ficheros del molde P4 y ficheros con firma del DINERO.
    #: Los dos se miden sobre TODOS los ficheros del kit, en alcance o no: el
    #: molde P4 vive entero fuera del alcance del molde ▸ (chef-privado: 0 de 9
    #: en ▸ y 11 hojas P4) y una tabla de recuento puede estar en un fichero
    #: que el motor no reconozca.
    p4_reconocidas, firmas_dinero = [], []
    zonas, bandas_neg = {}, {}
    candidatos = {'caja': [], 'cobros': [], 'negocio': [], 'areas': [],
                  'areas2': []}
    papeles, textos = {}, {}
    for fname in ficheros:
        wb = abrir(fname)
        recon = hojas_reconocidas(wb)
        horas[fname], cierres[fname] = [], []

        # --- CB-E4: ¿este fichero trae el «Cómo personalizar» heredado? -----
        if 'Instrucciones' in wb.sheetnames:
            ins = wb['Instrucciones']
            for r in range(1, ins.max_row + 1):
                for c in (2, 1):
                    v = ins.cell(row=r, column=c).value
                    if isinstance(v, str) and RX_PERSONALIZAR.match(
                            v.strip().rstrip(':')):
                        ctx['personalizar'][fname] = True
                        break
        ctx['papel_instrucciones'][fname] = papel_instrucciones(recon)

        if not marca_cb:
            for ws in wb.worksheets:
                if any('ChefBusiness' in c.value
                       for row in ws.iter_rows() for c in row
                       if isinstance(c.value, str)):
                    marca_cb = True
                    break

        # --- hora ancla: SÓLO de las hojas de apertura, estén o no en alcance
        for ws in wb.worksheets:
            if not ws.title.lower().startswith('apertura'):
                continue
            g = geometria(ws)
            col_h = _col_tiempo(g['cols']) if g else None
            if col_h:
                for r in range(g['hr'] + 1, (g['contador'] or ws.max_row)):
                    if not isinstance(ws.cell(row=r, column=1).value, int):
                        continue
                    v = ws.cell(row=r, column=col_h).value
                    if isinstance(v, str) and RX_HORA.match(v.strip()):
                        horas[fname].append(v.strip())
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str):
                    horas[fname] += [m.group(0)
                                     for m in RX_HORA_EN_TEXTO.finditer(v)]
        for ws in wb.worksheets:
            if not ws.title.lower().startswith('cierre'):
                continue
            g = geometria(ws)
            col_h = _col_tiempo(g['cols']) if g else None
            if not col_h:
                continue
            for r in range(g['hr'] + 1, (g['contador'] or ws.max_row)):
                if not isinstance(ws.cell(row=r, column=1).value, int):
                    continue
                v = ws.cell(row=r, column=col_h).value
                if isinstance(v, str) and v.strip() and not RX_HORA.match(
                        v.strip()):
                    cierres[fname].append(v.strip())

        # --- nombre del kit y sufijo de metadata, del TÍTULO del documento.
        # VOTAN TODOS los ficheros del kit (también los que el motor no
        # reconoce): antes sólo votaban los del molde ▸ y en catering el 08
        # empataba 1-1 con el 09 recién construido (crítico-5).
        # Es la única fuente estable: el motor reescribe las Instrucciones
        # (donde estaba «Kit de Tareas — <kit>») pero deja el patrón
        # «<algo> — <kit> · <sufijo>» en el título que él mismo compone.
        tit = wb.properties.title or ''
        if ' · ' in tit:
            cabeza, sufijo = tit.rsplit(' · ', 1)
            sufijos.append(sufijo)
            if ' — ' in cabeza:
                kits.append(cabeza.rsplit(' — ', 1)[1].strip())
        if not recon:
            # CB-E6 — un fichero fuera del molde ▸ todavía puede ser del molde
            # P4, que `aplicar` normaliza por su cuenta (DV, contador honesto,
            # CF y bio). Cuenta como molde RECONOCIDO: si no, chef-privado —11
            # hojas P4 y ni una ▸— abortaría por «molde desconocido» siendo el
            # kit que el motor mejor arregla de la sub-familia.
            if any(geometria_p4(ws) for ws in wb.worksheets
                   if ws.title != 'Instrucciones'):
                p4_reconocidas.append(fname)
            # CB-E7 — firma del dinero en un fichero que no está en alcance.
            if any(_tiene(wb, d) for d in
                   (fila_recuento, fila_registro_mensual, fila_liquidacion,
                    fila_registro_eventos)):
                firmas_dinero.append(fname)
            continue
        ctx['ficheros'].append(fname)
        if any(t == 'checklist' for t in recon.values()):
            ctx['con_checklist'].add(fname)
        # --- papel del fichero, por ESTRUCTURA (R3-a). Se ACUMULAN los
        #     candidatos y se decide al final: con `elif` sobre la marcha, el
        #     orden alfabético del listado elegía por nosotros.
        papel, detalle = papel_del_fichero(wb)
        if papel:
            candidatos[papel].append(fname)
            papeles[fname] = detalle
        # CB-E7 — firma del dinero de los ficheros EN alcance. Se lee de
        # `detalle` y no de `papel` porque un fichero puede traer la firma y
        # quedarse sin papel (`papel_del_fichero` mira las cuatro firmas del
        # dinero ANTES que nada, así que si devuelve 'negocio' o 'areas' es que
        # las cuatro son False).
        if any(detalle.get(k) for k in ('recuento', 'registro', 'liquidacion',
                                        'registro_eventos')):
            firmas_dinero.append(fname)
        # T-02 — el marco se lee con la MISMA normalización que se le va a
        # aplicar (`texto_tpv_caja` sólo en el fichero de caja): si no, la 1.ª
        # pasada compararía «Encender TPV / POS» y la 2.ª «Comprobar que el
        # TPV…», y `anotar_duplicados` daría dos resultados distintos.
        textos[fname] = tareas_del_libro(wb, caja=(papel == 'caja'),
                                         cobros=(papel == 'cobros'))
        # T-04 — nombres reales de zona y bandas reales del negocio.
        if papel in ('areas', 'areas2'):
            zonas[fname] = _zonas_del_fichero(wb)
        elif papel == 'negocio':
            bandas_neg[fname] = _bandas_del_fichero(wb)

        # --- papeles secundarios, también por ESTRUCTURA -------------------
        if any(t == 'calendario' for t in recon.values()):
            ctx['f_calendario'] = ctx['f_calendario'] or fname
        if any(t == 'briefing' for t in recon.values()):
            ctx['f_briefing'] = ctx['f_briefing'] or fname
        if sum(1 for t in recon if t in PLANTILLA_07) >= 2:
            ctx['f_plantilla'] = ctx['f_plantilla'] or fname
        ritmos = []
        for t, k in recon.items():
            if k != 'checklist' or t in PLANTILLA_07:
                continue
            g = geometria(wb[t])
            if g:
                ritmos.append(cadencia(wb[t], g))
        # TODAS las hojas del fichero, no la mayoría: 03 (manager) tiene una
        # diaria y un handover además de la semanal y la mensual, y como
        # «fichero de lo periódico» se llevaba la etiqueta que le toca al 05.
        if len(ritmos) >= 2 and all(x == 'Antelación' for x in ritmos):
            ctx['f_eventos'] = ctx['f_eventos'] or fname
        elif len(ritmos) >= 2 and all(x in ('Día', 'Cadencia') for x in ritmos):
            ctx['f_periodico'] = ctx['f_periodico'] or fname


    # --- CB-E6: ¿ha entendido el motor ALGO de este producto? -------------
    # Va antes que la resolución de papeles porque sin un solo molde
    # reconocido no hay papeles que resolver, y sobre todo antes de que
    # `main.py` empiece a guardar ficheros: la salida que esto impide es
    # «11 ficheros guardados, 0 fórmulas, censo en verde» sobre un kit que el
    # motor no ha leído. Los moldes que cuentan son los que el motor SABE
    # tratar hoy: ▸ (`geometria`), registro mensual, registro de eventos,
    # liquidación, calendario y briefing (los seis de `hojas_reconocidas`) y
    # P4 (`geometria_p4`). Cuando CB-E5 añada `geometria_plano`, el molde
    # PLANO se suma aquí y estos cuatro kits dejarán de abortar.
    if not ctx['ficheros'] and not p4_reconocidas:
        raise MoldeDesconocido(
            f'MOLDE DESCONOCIDO en {carpeta}: ninguno de los {len(ficheros)} '
            'ficheros del producto tiene una sola hoja que el motor sepa '
            'tratar (ni molde ▸, ni P4, ni registro mensual, ni registro de '
            'eventos, ni liquidación, ni calendario, ni briefing). El motor NO '
            'guarda un producto que no ha entendido: se abortaría dejándolo '
            'con la metadata de la v2.0 y el contenido de la v1.1. Ficheros: '
            f'{sorted(ficheros)}.')

    # --- resolución de papeles: uno y sólo uno por papel -----------------
    for papel in ('caja', 'cobros', 'negocio'):
        if len(candidatos[papel]) > 1:
            raise KitAmbiguo(
                f'DETECCIÓN AMBIGUA del fichero de {papel.upper()} en '
                f'{carpeta}: {candidatos[papel]} cumplen la misma firma '
                f'estructural ({ {f: papeles[f] for f in candidatos[papel]} }). '
                'El motor NO adivina: revisa las cabeceras antes de seguir.')
    areas = candidatos['areas'] or candidatos['areas2']
    if len(areas) > 1:
        raise KitAmbiguo(
            f'DETECCIÓN AMBIGUA del fichero de ÁREAS en {carpeta}: {areas}. '
            'El motor NO adivina.')
    # m6 — el fichero del dinero ocupa SIEMPRE la ranura `f_caja`, sea de
    # mostrador o de eventos: así las 20 referencias que ya existen (el bloque
    # «Se conecta con», `colapsar_duplicados`, `aplicar`, los gates) siguen
    # apuntando al mismo sitio y lo que cambia es el VOCABULARIO, no la
    # topología del kit. Lo que no puede haber es los dos a la vez: serían dos
    # ficheros del dinero en el mismo producto con dos modelos incompatibles, y
    # el motor no adivina cuál manda (R3-a).
    if candidatos['caja'] and candidatos['cobros']:
        raise KitAmbiguo(
            f'DETECCIÓN AMBIGUA del fichero del DINERO en {carpeta}: '
            f'{candidatos["caja"]} tiene arqueo de mostrador y '
            f'{candidatos["cobros"]} liquidación por evento. Un kit no puede '
            'entregar los dos modelos: revisa las cabeceras antes de seguir.')
    ctx['modelo_caja'] = 'eventos' if candidatos['cobros'] else 'mostrador'
    del_dinero = candidatos['caja'] or candidatos['cobros']
    ctx['f_caja'] = del_dinero[0] if del_dinero else None
    ctx['f_negocio'] = (candidatos['negocio'][0] if candidatos['negocio']
                        else None)
    ctx['f_areas'] = areas[0] if areas else None
    ctx['papeles'] = papeles
    ctx['areas_nombres'] = zonas.get(ctx['f_areas'], [])
    ctx['negocio_bandas'] = bandas_neg.get(ctx['f_negocio'], [])
    # CB-E7 — kit sin negocio y sin caja. No basta con que las dos ranuras
    # estén vacías: se exige además que NINGÚN fichero del kit traiga firma de
    # recuento, registro mensual, liquidación o registro de eventos. Si alguno
    # la trajera, el kit SÍ tiene fichero del dinero y lo que hay es un fallo
    # de detección —que es justo lo que el gate `negocio_precargado` debe
    # seguir cantando en rojo—, no una topología distinta.
    ctx['sin_caja'] = (ctx['f_negocio'] is None and ctx['f_caja'] is None
                       and not firmas_dinero)
    ctx['firmas_dinero'] = sorted(set(firmas_dinero))
    ctx['ficheros_p4'] = sorted(p4_reconocidas)

    # La hora ancla y el literal de cierre NO pueden salir de los ficheros de
    # caja y de negocio: son justo los que el motor PRECARGA con esos valores.
    # Si salieran de ahí, cada pasada leería lo que escribió la anterior y el
    # ancla se desplazaría 15 minutos por ejecución (medido: 07:00 → 06:45).
    propios = [f for f in ficheros
               if f not in (ctx['f_caja'], ctx['f_negocio'])]
    reales = sorted(h for f in propios for h in horas.get(f, [])
                    if RX_HORA.match(h))
    if reales:
        ctx['hora_apertura'] = reales[0]
    lit = [c for f in propios for c in cierres.get(f, [])]
    if lit:
        ctx['literal_cierre'] = _mayoria(lit)
    ctx['kit'] = _mayoria(kits)
    ctx['sufijo'] = _mayoria(sufijos)
    # §7-bis.17 — el pie CONSERVA la marca que el producto ya tiene. Hasta esta
    # tanda la línea era el literal de AI Chef Pro y `reescribir_instrucciones`
    # la escribía en los 11 ficheros, así que el dry-run de sushi-bar convertía
    # «— Kit de Tareas: Sushi Bar · ChefBusiness Consultoria Gastronomica ·
    # chefbusiness.co» en «— Kit de Tareas Recurrentes · Sushi Bar · AI Chef
    # Pro · aichef.pro»: un REBRANDEO de seis productos, que es exactamente lo
    # que la decisión firmada prohíbe (y que ningún gate miraba). La estructura
    # sí se unifica con la de la familia; lo que se preserva es la marca.
    # chef-privado no menciona ChefBusiness en ninguna celda y se queda, como
    # está firmado, con AI Chef Pro.
    ctx['marca'], ctx['dominio'] = (
        ('ChefBusiness Consultoría Gastronómica', 'chefbusiness.co')
        if marca_cb else ('AI Chef Pro', 'aichef.pro'))
    ctx['pie'] = ('— Kit de Tareas Recurrentes · '
                  + (ctx['kit'] + ' · ' if ctx['kit'] else '')
                  + ctx['marca'] + ' · ' + ctx['dominio'])
    if producto:
        ctx['producto'] = producto
    CTX.clear()
    CTX.update(ctx)
    # R3-f — los textos del MARCO (08 y 09) viven fuera de `ctx` a propósito:
    # `main.py` vuelca el contexto entero en el informe y esto son cientos de
    # frases que no aportan nada allí.
    CTX['tareas_marco'] = {f: textos.get(f, [])
                           for f in (ctx['f_negocio'], ctx['f_caja']) if f}
    return ctx


def tareas_del_libro(wb, caja=False, cobros=False):
    """Todos los textos de la columna «Tarea» de un libro, en forma ESTABLE.

    «Estable» = ya pasados por las normalizaciones de texto del motor, que son
    idempotentes: así la 1.ª pasada (que los lee crudos) y la 2.ª (que los lee
    ya reescritos) comparan exactamente lo mismo y `anotar_duplicados` no puede
    dar un resultado distinto en cada pasada. `caja=True` añade la de T-02, que
    sólo se aplica al fichero de caja de MOSTRADOR; `cobros=True` quita la de
    DOM-01 (`texto_facturado`), que reescribe una cuenta de arqueo que en el
    modelo por eventos no existe (m6).

    Los dos parámetros se pasan explícitamente y NO se leen de `CTX`: esta
    función corre DENTRO de `contexto()`, es decir antes de que `CTX` se
    actualice, y leería el contexto del kit ANTERIOR.
    """
    fuera = []
    for ws in wb.worksheets:
        g = geometria(ws) or geometria_p4(ws)
        if not g:
            continue
        for r in range(g['hr'] + 1, (g.get('contador') or ws.max_row)):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, str) and v.strip() and v.strip() != 'Tarea':
                fuera.append(forma_estable(texto_tpv_caja(v) if caja else v,
                                           facturado=not cobros))
    return fuera


#: T-04 — «Apertura Cocina» / «Cierre Sala» → «cocina», «sala». Si la hoja no
#: lleva zona en el título («Apertura» a secas, «Producción»), no se deriva
#: nada: el paréntesis se OMITE antes que inventarse una zona.
RX_CICLO_ZONA = re.compile(
    r'(?i)^\s*(?:pre[\s-]?apertura|post[\s-]?cierre|apertura|cierre)\s+'
    r'(?:de\s+(?:la|el|los|las)\s+|del\s+|de\s+)?')


def _zonas_del_fichero(wb):
    """Nombres reales de zona del fichero de ÁREAS, en orden de hoja."""
    fuera = []
    for ws in wb.worksheets:
        if not geometria(ws):
            continue
        nombre = RX_CICLO_ZONA.sub('', ws.title).strip()
        if not nombre or nombre.lower() == ws.title.strip().lower():
            continue                     # la hoja no nombra ninguna zona
        clave = nombre.lower()
        if clave not in fuera:
            fuera.append(clave)
    return fuera


def _bandas_del_fichero(wb):
    """Rótulos reales de las bandas de sección del fichero de NEGOCIO.

    Verificado el 2026-08-23 en las 11 carpetas de la familia: los 08/18/10 se
    entregan como listas PLANAS, sin ninguna banda. Por eso el paréntesis de
    «(accesos, luces, clima, terraza)» no se puede derivar y se omite: era el
    único sitio del kit que afirmaba que el local tiene terraza.
    """
    fuera = []
    for ws in wb.worksheets:
        g = geometria(ws)
        if not g:
            continue
        for r in range(g['hr'] + 1, (g['contador'] or ws.max_row)):
            if not es_fila_seccion(ws, r):
                continue
            v = ws.cell(row=r, column=1).value
            if not isinstance(v, str):
                continue
            nombre = RX_NOTA_DUP.split(_nombre_seccion(v))[0].strip()
            clave = nombre.lower().strip('→- ')
            if clave and clave not in fuera:
                fuera.append(clave)
    return fuera


def parentesis(items, maximo=4):
    """« (a, b, c)» o cadena vacía si no hay nada que enumerar (T-04)."""
    items = [i for i in (items or []) if i]
    if not items:
        return ''
    return ' (' + ', '.join(items[:maximo]) + ')'


def _tiene(wb, detector):
    return any(detector(ws) for ws in wb.worksheets)


#: CB-E1 (efecto colateral obligatorio) — los rótulos del molde se comparan SIN
#: TILDE. `ortografia()` corre al final de la pasada y reescribe la cabecera:
#: con una comparación literal, la 1.ª pasada no vería la columna de tiempo de
#: sushi-bar («Hora Limite») y la 2.ª —ya corregida— sí, con cabecera y
#: subtítulo reescritos y la columna pintada de verde SÓLO en la segunda. El
#: gate de idempotencia se pondría rojo sin que nada estuviera roto. En los 13
#: kits de la familia, que ya escriben las tildes, esto no cambia nada:
#: verificado con `regresion.py` sobre cafetería y hotel (0 diferencias).
def _rotulo(nombre, candidatos):
    """Devuelve la clave de `candidatos` que es `nombre` salvo tildes, o None."""
    objetivo = _sin_tildes(nombre)
    for clave in candidatos:
        if _sin_tildes(clave) == objetivo:
            return clave
    return None


def _col_tiempo(cols):
    for nombre in ('Hora Límite', 'Hora', 'Día', 'Cadencia', 'Antelación',
                   'Cuándo'):
        clave = _rotulo(nombre, cols)
        if clave is not None:
            return cols[clave]
    return None


def _sumar_minutos(hhmm, minutos):
    h, m = (int(x) for x in hhmm.split(':'))
    t = (h * 60 + m + minutos) % (24 * 60)
    return f'{t // 60:02d}:{t % 60:02d}'


# ==========================================================================
# §1.5 / §2 — normalización del checklist
# ==========================================================================
EDITABLES = {'Responsable', 'Zona', 'Puesto', 'Hora Límite', 'Hora', 'Día',
             'Cadencia', 'Antelación', 'Cuándo', CAB_MARCA, 'Firma', 'Notas'}


def crear_contador(ws, fname, cambios):
    """CB-E2 — crea la fila de totales en la hoja de checklist que no la tiene.

    `geometria` devuelve geometría válida y `contador = None` en
    `09-plantilla-personalizable.xlsx:'Plantilla en Blanco'` de sushi-bar y de
    asador: la hoja tiene cabecera, 15 filas numeradas y un pie «Verificado
    por:», pero ninguna fila de totales. El motor sólo REESCRIBE contadores, no
    los crea, así que la hoja entraba en alcance con 0 fórmulas mientras sus
    «Instrucciones» estrenaban los bloques «Cómo cuenta el contador» y «Filas
    libres»: tres bloques prometiendo lo que la hoja no hace (§0.2 defecto 3).

    Se limita a escribir el RÓTULO en una fila nueva; el resto lo hace la
    maquinaria de siempre, porque a partir de ahí `geometria` ya encuentra el
    contador: `normalizar_checklist` mete las 5 filas libres dentro del rango y
    `_contador` escribe el COUNTIFS y el denominador honesto. Escribir aquí las
    fórmulas duplicaría ese código y las dos copias podrían divergir.

    NO se aplica al fichero de NEGOCIO ni al de CAJA, y no es una precaución
    teórica: `kit-tareas-pasteleria` —un kit LIVE— tiene exactamente esta firma
    en `09-apertura-cierre-caja.xlsx:'Apertura de Caja'` y `'Cierre de Caja'`,
    donde el total no es un recuento de tareas sino el Resumen de Cierre del
    arqueo. Crearles un contador habría cambiado un producto publicado.
    """
    if fname in (CTX.get('f_caja'), CTX.get('f_negocio')):
        return False
    g = geometria(ws)
    if not g or g['contador'] is not None:
        return False
    ncol = max(ws.max_column, max(g['cols'].values()))
    # última fila ESCRITA del cuerpo (no la última numerada): si hubiera una
    # banda de sección por debajo, el contador tiene que ir detrás de ella.
    cuerpo_fin = g['ultima']
    for r in range(g['ultima'] + 1, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None
               for c in range(1, ncol + 1)):
            break
        cuerpo_fin = r
    fila = g['ultima'] + 2                 # una fila en blanco de separación
    insertar_filas(ws, fila, 1)
    cel = ws.cell(row=fila, column=1, value=ETIQ_CONTADOR)
    cel._style = copy.copy(ws.cell(row=g['hr'], column=1)._style)
    cambios.append(f'CB-E2 «{ws.title}»: fila de totales CREADA (la hoja '
                   'entraba en alcance sin contador y sus Instrucciones ya '
                   'explicaban uno)')
    return True


def normalizar_checklist(ws, cambios):
    """DV «✓,—,N/A», 5 filas libres DENTRO del rango, contador honesto, verdes y
    formato condicional. §2.1 + §2.2 + §1.5."""
    g = geometria(ws)
    if not g:
        return None
    hr, marca, contador, ultima = g['hr'], g['marca'], g['contador'], g['ultima']
    ncol = max(ws.max_column, max(g['cols'].values()))

    # --- 5 filas libres con formato, DV y CF dentro del rango contado -----
    if contador:
        # El ancla NO es la última tarea numerada sino la última fila con algo
        # escrito: cuando `colapsar_duplicados` reduce el ÚLTIMO bloque a una
        # banda de remisión (COM-R2-07), esa banda queda por debajo de la
        # última tarea y las filas libres caían encima de ella — sobre una
        # celda combinada, que en openpyxl es de sólo lectura.
        cuerpo_fin = ultima
        for r in range(ultima + 1, contador):
            if any(ws.cell(row=r, column=c).value is not None
                   for c in range(1, ncol + 1)):
                cuerpo_fin = r
        hueco = contador - 1 - cuerpo_fin
        faltan = (HOLGURA + 1) - hueco          # +1 = fila en blanco separadora
        if faltan > 0:
            insertar_filas(ws, cuerpo_fin + 1, faltan)
            contador += faltan
            cambios.append(f'«{ws.title}»: +{faltan} filas libres dentro del '
                           f'rango contado (total {HOLGURA})')
        fin = cuerpo_fin + HOLGURA
        for r in range(cuerpo_fin + 1, fin + 1):
            _desmerge_fila(ws, r)
            for c in range(1, ncol + 1):
                cel = ws.cell(row=r, column=c)
                cel.value = None
                cel._style = copy.copy(ws.cell(row=ultima, column=c)._style)
                # DOM-R2-23: la columna «Nº» también va verde. Numerarlas NO
                # vale: `geometria` mide la última tarea por el entero de la
                # columna A y las filas libres pasarían a contar como tareas,
                # así que cada pasada añadiría 5 filas más.
                _verde(cel)
    else:
        fin = ultima

    letra = get_column_letter(marca)

    # --- verdes: columnas editables de las filas de tarea (TEC-15) --------
    for r in range(hr + 1, ultima + 1):
        if es_fila_seccion(ws, r):
            for c in range(1, ncol + 1):
                cel = ws.cell(row=r, column=c)
                if _es_color(cel, VERDE):
                    _relleno(cel, SECCION)      # el verde sólo = editable
            continue
        if not isinstance(ws.cell(row=r, column=1).value, int):
            continue
        for nombre, c in g['cols'].items():
            # CB-E1: sin tildes. «Hora Limite» de sushi-bar y asador NO estaba
            # en EDITABLES y su columna se publicaba en blanco mientras las
            # Instrucciones prometían «las celdas verdes son editables».
            if _sin_tildes(nombre) in {_sin_tildes(e) for e in EDITABLES}:
                _verde(ws.cell(row=r, column=c))

    # --- validación de datos ---------------------------------------------
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if dv.type != 'list']
    dv = DataValidation(
        type='list', formula1=DV_LISTA, allow_blank=True,
        showErrorMessage=True, errorStyle='stop',
        errorTitle=DV_ERROR_TIT, error=DV_ERROR)
    ws.add_data_validation(dv)
    for r in range(hr + 1, fin + 1):
        if es_fila_seccion(ws, r):
            continue
        if isinstance(ws.cell(row=r, column=1).value, int) or ultima < r <= fin:
            dv.add(ws.cell(row=r, column=marca))

    # --- formato condicional: fila verde al marcar ✓ ----------------------
    _ok = verde_ok()
    ws.conditional_formatting.add(
        f'A{hr + 1}:{get_column_letter(ncol)}{fin}',
        FormulaRule(formula=[f'${letra}{hr + 1}="{MARCA_OK}"'],
                    fill=PatternFill('solid', start_color=_ok,
                                     end_color=_ok)))

    # --- contador honesto (§2.1) ------------------------------------------
    if contador:
        _contador(ws, contador, hr, fin, marca, ncol, cambios)
    return {'hr': hr, 'fin': fin, 'marca': marca, 'contador': contador,
            'ultima': ultima}


def capacidad(ws, col, size=10):
    """Caracteres que caben en UNA línea de `col` (heurística del avance del 0)."""
    ancho = ws.column_dimensions[get_column_letter(col)].width or 8.43
    return max(4, int(ancho * 11 / (size or 11)))


def autoalto(ws, fila, col):
    """TEC-R2-13/TEC-R2-04 — quita la altura fija a la fila cuyo texto no cabe.

    Con `wrap_text` y altura fija de 24 pt sólo se ven ~1,6 líneas de 10 pt: los
    textos de 58-73 caracteres en una columna de 50 se publican RECORTADOS por
    el final, que es donde está el dato. Dejando `height = None` (sin
    `customHeight`) Excel autoajusta al abrir. Es la razón por la que la fila
    de 0,02 €, la única sin altura fija, era la única moneda que se leía entera.
    """
    cel = ws.cell(row=fila, column=col)
    if not isinstance(cel.value, str) or ws.row_dimensions[fila].height is None:
        return False
    if not (cel.alignment and cel.alignment.wrap_text):
        return False
    if len(cel.value) <= capacidad(ws, col, cel.font.size or 11):
        return False
    ws.row_dimensions[fila].height = None
    return True


def autoaltos(ws, cambios):
    """R3-d — la pasada de alturas va DESPUÉS de todos los cambios de texto.

    Vivía dentro de `normalizar_checklist`, que corre ANTES de
    `textos_de_tarea`: la 1.ª pasada medía el texto CORTO (original) y lo daba
    por bueno, y la 2.ª medía el largo (ya con «(refrigeración 0-4 °C) — anota
    la lectura: ____ °C») y le quitaba la altura fija. Una fila de diferencia
    entre pasadas — heladería, 08-apertura-cierre-negocio.xlsx:'Apertura del
    Negocio'!21 — y el gate de idempotencia en rojo sin nada realmente roto.
    Midiendo al final, las dos pasadas leen el MISMO texto.
    """
    g = geometria(ws)
    if not g:
        return
    n = 0
    for r in range(g['hr'] + 1, g['ultima'] + 1):
        if es_fila_seccion(ws, r):
            continue
        if not isinstance(ws.cell(row=r, column=1).value, int):
            continue
        if autoalto(ws, r, 2):
            n += 1
    if n:
        cambios.append(f'«{ws.title}»: {n} filas sin altura fija (el texto no '
                       'cabía y se imprimía recortado)')


def legibilidad_tareas(ws, cambios):
    """CB-E9 (a) — el texto de la tarea, legible impreso: `wrap_text` y ancho.

    Los siete kits de CB traen la columna «Tarea» a 48 unidades y SIN
    `wrap_text` en ninguna celda: 107 tareas de sushi-bar y 129 de asador miden
    más de 48 caracteres, así que o invaden la columna «Zona» (si está vacía) o
    se cortan (si no lo está). Se arregla con las dos cosas a la vez: se
    envuelve el texto —que es lo que la familia ▸ ya hace: cafetería trae
    `wrap_text` en las 446 celdas equivalentes— y se ensancha la columna con lo
    que sobre del presupuesto de A4 apaisado, hasta `ANCHO_TAREA_CB`.

    La altura se deja en `None` a propósito (Excel autoajusta al abrir): es la
    misma decisión que documenta `autoalto`, y una altura fija de 24 pt con
    `wrap_text` enseña ~1,6 líneas de 10 pt, o sea recorta por el final.

    Sólo corre en la sub-familia CB. La firma del defecto es IDÉNTICA en cuatro
    kits LIVE —cafetería (446 celdas), pizzería (319), hamburguesería (292) y
    dark-kitchen (275)—, así que una versión sin `sub_cb()` habría reescrito
    cuatro productos publicados que §7-bis.24 congela.
    """
    if not sub_cb():
        return
    g = geometria(ws)
    if not g:
        return
    n = 0
    for r in range(g['hr'] + 1, (g['contador'] or ws.max_row) + 1):
        if es_fila_seccion(ws, r):
            continue
        cel = ws.cell(row=r, column=2)
        if cel.alignment and cel.alignment.wrap_text:
            continue
        a = cel.alignment
        cel.alignment = Alignment(wrap_text=True, vertical='top',
                                  horizontal=a.horizontal if a else None,
                                  indent=a.indent if a else 0)
        n += 1
    letra = get_column_letter(2)
    actual = ws.column_dimensions[letra].width or 8.43
    margen = ANCHO_A4 - ancho_util(ws)
    nuevo = min(ANCHO_TAREA_CB, actual + max(margen, 0))
    if nuevo > actual + 0.5:
        ws.column_dimensions[letra].width = round(nuevo)
        cambios.append(f'CB-E9 «{ws.title}»: columna «Tarea» de {actual:g} a '
                       f'{ws.column_dimensions[letra].width:g} unidades '
                       f'(la fila cabe en A4 apaisado: '
                       f'{ancho_util(ws):.0f} ≤ {ANCHO_A4})')
    if n:
        cambios.append(f'CB-E9 «{ws.title}»: {n} celdas de tarea con ajuste de '
                       'texto (el texto largo se imprimía cortado o pisando '
                       '«Zona»)')


def _es_color(cel, color):
    f = cel.fill
    return (f is not None and f.fill_type == 'solid' and f.fgColor is not None
            and isinstance(f.fgColor.rgb, str)
            and f.fgColor.rgb.upper().endswith(color))


def _contador(ws, fila, hr, fin, marca, ncol, cambios):
    """«Tareas completadas: X de Y» agrupado bajo la columna que mide (TEC-21).

    Molde de 01-07: etiqueta combinada de A hasta `marca-3`, numerador en
    `marca-2`, «de» en `marca-1` y denominador en `marca`. 08 y 09 lo tenían
    una columna a la izquierda, así que el «de» caía bajo «Responsable».
    """
    letra = get_column_letter(marca)
    tarea = get_column_letter(2)
    # TEC-R2-09: el numerador exige TEXTO en la tarea. Con `COUNTIF(F,"✓")` a
    # secas, un ✓ en una de las 5 filas libres vacías daba «32 de 30».
    num = (f'=COUNTIFS({tarea}{hr + 1}:{tarea}{fin},"?*",'
           f'{letra}{hr + 1}:{letra}{fin},"{MARCA_OK}")')
    # DOM-R2-02: del denominador sale N/A («no aplica en este local») y NADA
    # más. «—» = no hecho: si también saliera, el turno que se salta tareas las
    # marca «—» y la hoja que el kit manda archivar como prueba imprime 100 %.
    den = (f'=COUNTIF({tarea}{hr + 1}:{tarea}{fin},"?*")'
           f'-COUNTIF({letra}{hr + 1}:{letra}{fin},"N/A")')

    _desmerge_fila(ws, fila)
    estilo = None
    for c in range(1, ncol + 1):
        cel = ws.cell(row=fila, column=c)
        if isinstance(cel.value, str) and RX_CONTADOR.match(cel.value) \
                and estilo is None:
            estilo = copy.copy(cel._style)
        cel.value = None
    fin_etq = marca - 3
    cel = ws.cell(row=fila, column=1, value=ETIQ_CONTADOR)
    if estilo is not None:
        cel._style = estilo
    if fin_etq >= 2:
        _merge(ws, f'A{fila}:{get_column_letter(fin_etq)}{fila}')
        cambios.append(f'«{ws.title}»: contador agrupado '
                       f'(etiqueta A:{get_column_letter(fin_etq)}, '
                       f'«x de y» bajo «{CAB_MARCA}»)')
    ws.cell(row=fila, column=marca - 2).value = num
    reg(ws, ws.cell(row=fila, column=marca - 2).coordinate, num)
    ws.cell(row=fila, column=marca - 1).value = 'de'
    ws.cell(row=fila, column=marca).value = den
    reg(ws, ws.cell(row=fila, column=marca).coordinate, den)


# ==========================================================================
# §2.4 — cabecera de la columna de tiempo y subtítulo de la fila 2
# ==========================================================================
SUBTITULO = {
    'Día': 'Semana del ___/___/______ al ___/___/______    '
           'Responsable: _________________________',
    'Cadencia': 'Mes: ________________   Año: ______    '
                'Responsable: _________________________',
    'Antelación': 'Evento / temporada: ___________________    '
                  'Fecha: ___/___/______    '
                  'Responsable: _________________________',
}


def cadencia(ws, g):
    """Clasifica la columna de tiempo por su CONTENIDO (DOM-25/TEC-10).

    DOM-R2-24: «Hora Límite» sólo se conserva cuando TODOS los valores son
    horas. Si hay hitos que no son hora («Cierre», «Servicio», «Cambio», «Si
    aplica», «Según entrega») y no dominan Día/Cadencia/Antelación, la columna
    pasa a titularse «Cuándo» — impreso, «Hora Límite: Si aplica» no dice nada.
    """
    col = _col_tiempo(g['cols'])
    if not col:
        return None
    votos = {'hora': 0, 'Día': 0, 'Cadencia': 0, 'Antelación': 0, 'otro': 0}
    for r in range(g['hr'] + 1, g['ultima'] + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str) or not v.strip():
            continue
        v = v.strip()
        if RX_HORA.match(v):
            votos['hora'] += 1
        elif (v.lower() in DIAS or v.lower().startswith(DIAS)
                or RX_DIA_N.match(v) or RX_FINDE.match(v)):
            votos['Día'] += 1
        elif RX_CADENCIA.match(v):
            votos['Cadencia'] += 1
        elif RX_ANTELACION.search(v):
            votos['Antelación'] += 1
        else:
            votos['otro'] += 1
    no_hora = sum(v for k, v in votos.items() if k != 'hora')
    if not no_hora:
        return None                        # todo son horas: «Hora Límite»
    gana = max(('Día', 'Cadencia', 'Antelación'), key=lambda k: votos[k])
    if votos[gana] * 2 >= no_hora and votos[gana]:
        return gana
    return 'Cuándo'


def ajustar_cabecera_tiempo(ws, cambios):
    g = geometria(ws)
    if not g:
        return
    nueva = cadencia(ws, g)
    if not nueva:
        return
    col = _col_tiempo(g['cols'])
    actual = ws.cell(row=g['hr'], column=col).value
    if actual != nueva:
        ws.cell(row=g['hr'], column=col).value = nueva
        cambios.append(f'«{ws.title}»: cabecera «{actual}» → «{nueva}» '
                       '(la columna no lleva horas)')
    sub = SUBTITULO.get(nueva)
    if sub and ws.cell(row=2, column=1).value != sub:
        ws.cell(row=2, column=1).value = sub
        cambios.append(f'«{ws.title}»: fila 2 acorde a la cadencia')


# ==========================================================================
# §2.9 — temperaturas con objetivo · §2.5 — referencias al Pack APPCC
# ==========================================================================
LECTURA = ' — anota la lectura: ____ °C'
OBJ_FRIO = ' (refrigeración 0-4 °C)'
OBJ_CONG = ' (congelación ≤ −18 °C)'
OBJ_AMBOS = ' (refrigeración 0-4 °C / congelación ≤ −18 °C)'

# La regla SÓLO pica en tareas de FRÍO: «Encender hornos y precalentar a
# temperatura de trabajo» casaba con un «temperatura» a secas y se le añadía
# «(refrigeración 0-4 °C / congelación ≤ −18 °C)» a un horno. Hacen falta las
# tres cosas: un verbo de registro, la palabra temperatura y un equipo de frío.
RX_TEMP = re.compile(r'(?i)\btemperaturas?\b')
RX_VERBO_TEMP = re.compile(
    r'(?i)\b(registrar|registro|anotar|verificar|comprobar|controlar|tomar|'
    r'medir|revisar)\b')
RX_EQUIPO_FRIO = re.compile(
    r'(?i)(c[áa]maras?|congelador|congelaci[óo]n|arc[óo]n|refrigera|nevera|'
    r'abatidor|vitrina|expositor|mural)')
RX_COLA_FRIO = re.compile(r'\s*\((refrigeraci[óo]n|congelaci[óo]n)\)\s*$',
                          re.I)
RX_APPCC = re.compile(r'APPCC')
# DOM-R2-09: la columna «Notas» sólo existe en los ficheros de negocio y de
# caja. En 01-07 el texto mandaba dejar constancia de una lectura —el registro
# que pide un inspector— en una columna que no está en la hoja.
SUST_APPCC = [
    (re.compile(r'^Registrar en hoja de control de temperaturas APPCC$'),
     'Si tienes el Pack APPCC, registra la lectura en su hoja de '
     'temperaturas; si no, anótala aquí mismo: ____ °C'),
    (re.compile(r'^Revisión de registros APPCC de la semana$'),
     'Revisar los registros de higiene de la semana: si tienes el Pack '
     'APPCC, en sus hojas; si no, en las tuyas'),
    (re.compile(r'^▸ Archiva las hojas firmadas junto con los registros '
                r'APPCC$'),
     '▸ Archiva las hojas firmadas junto a tus registros de higiene: si '
     'tienes el Pack APPCC, con sus hojas; si no, con los tuyos'),
]


#: DOM-01 (texto) — la fórmula del Cierre de Caja descuenta el fondo desde la
#: v2.0, pero la TAREA que manda calcularlo seguía dictando la cuenta vieja.
#: El operario que hace caso al texto y no a la hoja se lleva el fondo a las
#: ventas todos los días.
RX_FACTURADO = re.compile(
    r'(?i)total\s+facturado\s*=\s*efectivo\s*\+\s*tarjetas\s*\+\s*otros')
TXT_FACTURADO = 'Total facturado = efectivo contado − fondo + tarjetas + otros'


def texto_facturado(v):
    if not isinstance(v, str) or '=' not in v:
        return v
    return RX_FACTURADO.sub(TXT_FACTURADO, v)


def texto_tpv_caja(v):
    """T-02 — en el fichero de CAJA el TPV no se enciende: se COMPRUEBA.

    Idempotente: el texto de salida no empieza por «Encender», así que la 2.ª
    pasada no vuelve a picar. Se aplica SÓLO al fichero de caja (`aplicar`) y a
    la lectura del marco de `tareas_del_libro`, para que `anotar_duplicados`
    compare lo mismo en las dos pasadas.
    """
    if not isinstance(v, str) or not RX_TPV_CAJA.match(v):
        return v
    return TXT_TPV_CAJA


def tpv_de_caja(ws, cambios):
    """T-02 sobre una hoja de checklist del fichero de CAJA."""
    g = geometria(ws)
    if not g:
        return 0
    n = 0
    for r in range(g['hr'] + 1, (g['contador'] or ws.max_row)):
        cel = ws.cell(row=r, column=2)
        nuevo = texto_tpv_caja(cel.value)
        if nuevo != cel.value:
            cambios.append(f'«{ws.title}»: B{r} «{cel.value}» → '
                           f'«{TXT_TPV_CAJA}» (T-02: el hito de encender el '
                           'TPV es del fichero de negocio; aquí se comprueba)')
            cel.value = nuevo
            n += 1
    return n


def texto_grados(v):
    """DOM-R2-22 — «-18°C» / «>65°C» → «−18 °C» / «>65 °C» en todo el corpus."""
    if not isinstance(v, str) or '°' not in v:
        return v
    return RX_MENOS_RANGO.sub('−', RX_MENOS.sub('−', RX_GRADOS.sub(' °C', v)))


def texto_appcc(v):
    if not isinstance(v, str) or not RX_APPCC.search(v):
        return v
    if 'si tienes el Pack APPCC' in v or 'Si tienes el Pack APPCC' in v:
        return v
    for rx, bueno in SUST_APPCC:
        if rx.match(v.strip()):
            return bueno
    return v + ' (si tienes el Pack APPCC, regístralo allí; si no, anótalo ' \
               'junto a la tarea)'


def texto_temperatura(v):
    """§2.9 — objetivo y hueco para la lectura en las tareas de temperatura."""
    if not isinstance(v, str) or '____ °C' in v or RX_APPCC.search(v):
        return v
    if not (RX_TEMP.search(v) and RX_VERBO_TEMP.search(v)
            and RX_EQUIPO_FRIO.search(v)):
        return v
    # «Registrar temperatura cámara 1 (refrigeración)» ya trae la cola entre
    # paréntesis: se sustituye, no se encadena una segunda.
    v = RX_COLA_FRIO.sub('', v)
    bajo = v.lower()
    frio = bool(re.search(r'(?i)(c[áa]mara|refrigera|nevera|vitrina|'
                          r'expositor|mural|abatidor)', bajo))
    cong = bool(re.search(r'(?i)(congelad|congelaci|arc[óo]n)', bajo))
    obj = OBJ_AMBOS if frio and cong else (OBJ_CONG if cong else OBJ_FRIO)
    if '°C' in v:                      # el texto ya da su propia cifra
        obj = ''
    return v + obj + LECTURA


def forma_estable(v, facturado=True):
    """El texto tal y como quedará tras el motor (las cuatro son idempotentes).

    m6 — `facturado=False` en el fichero de COBROS: `texto_facturado` reescribe
    «Total facturado = efectivo + tarjetas + otros» a la cuenta del arqueo con
    fondo, y en un kit que factura por evento esa cuenta no existe.
    """
    if facturado:
        v = texto_facturado(v)
    return texto_temperatura(texto_grados(texto_appcc(v)))


def textos_de_tarea(ws, cambios, col_tarea=None, facturado=True):
    """§2.5 (Pack APPCC) en toda la hoja y §2.9 sólo en la columna «Tarea»."""
    n = 0
    for row in ws.iter_rows():
        for c in row:
            if not isinstance(c.value, str):
                continue
            nuevo = texto_grados(texto_appcc(
                texto_facturado(c.value) if facturado else c.value))
            if col_tarea and c.column == col_tarea:
                nuevo = texto_temperatura(nuevo)
            if nuevo != c.value:
                c.value = nuevo
                n += 1
    if n:
        cambios.append(f'«{ws.title}»: {n} textos con objetivo de temperatura '
                       'o referencia honesta al Pack APPCC')


# ==========================================================================
# §2.5 — bloques de 01 que duplican 08 y 09
# ==========================================================================
#: {rótulo de la banda: (claves de CTX, plantilla de la línea de remisión)}.
#: DOM-R2-07: el bloque SISTEMAS de la apertura duplicaba las 5 tareas de 08 y
#: 09 con TRES horas distintas para el mismo TPV (06:45, 07:15 y 11:30) bajo una
#: línea impresa que promete «No se duplican».
COLAPSO = {
    'CIERRE DE CAJA': (('f_caja',),
                       '  CIERRE DE CAJA → ver {0}: fondo, recuento por '
                       'denominaciones, Z del TPV y descuadre'),
    'CIERRE GENERAL': (('f_negocio',),
                       '  CIERRE GENERAL → ver {0}: alarma, luces, '
                       'climatización, cierre y llaves'),
    'SISTEMAS': (('f_negocio', 'f_caja'),
                 '  SISTEMAS → ver {0} (TPV, datáfono y música del local) y '
                 '{1} (fondo de caja, cajón y rollo de ticket)'),
}
#: m6 — la misma remisión con el vocabulario del modelo POR EVENTOS. Remitir a
#: «fondo, recuento por denominaciones, Z del TPV y descuadre» en un kit cuyo
#: fichero del dinero no tiene ninguna de esas cuatro cosas mandaría al lector a
#: buscar una hoja que no existe.
COLAPSO_EVENTOS = {
    'CIERRE DE CAJA': (('f_caja',),
                       '  CIERRE DE CAJA → ver {0}: liquidación del evento, '
                       'factura, anticipo y saldo pendiente'),
    'SISTEMAS': (('f_negocio', 'f_caja'),
                 '  SISTEMAS → ver {0} (TPV, datáfono y música del local) y '
                 '{1} (facturación y cobro de cada evento)'),
}


def colapso():
    """Plantillas de remisión del modelo de caja que tiene este kit (m6)."""
    if not es_modelo_eventos():
        return COLAPSO
    fuera = dict(COLAPSO)
    fuera.update(COLAPSO_EVENTOS)
    return fuera


def _filas_del_bloque(ws, banda, tope):
    filas = []
    for rr in range(banda + 1, tope):
        if es_fila_seccion(ws, rr):
            break
        if isinstance(ws.cell(row=rr, column=1).value, int):
            filas.append(rr)
        elif filas:
            break
    return filas


def colapsar_duplicados(ws, cambios):
    """Reduce a una línea los bloques que repiten 08 y 09 (DOM-09/TEC-14).

    COM-R2-07: la remisión va en la BANDA de sección (columna A, combinada), no
    en una fila numerada. Como fila de tarea ocupaba un número, entraba en el
    denominador del contador («de 12» con 10 tareas ejecutables) y salía sin
    Responsable ni Hora, justo lo que la tarjeta de venta promete que no pasa.
    """
    hecho = True
    while hecho:
        hecho = False
        g = geometria(ws)
        if not g:
            return
        tope = g['contador'] or ws.max_row
        for r in range(g['hr'] + 1, tope):
            if not es_fila_seccion(ws, r):
                continue
            titulo = (ws.cell(row=r, column=1).value or '').strip().upper()
            tabla = colapso()
            if titulo not in tabla:
                continue
            claves, plantilla = tabla[titulo]
            destinos = [CTX.get(k) for k in claves]
            if not all(destinos):
                continue
            filas = _filas_del_bloque(ws, r, tope)
            if not filas:
                continue
            ws.cell(row=r, column=1).value = plantilla.format(*destinos)
            eliminar_filas(ws, filas[0], len(filas))
            cambios.append(f'«{ws.title}»: bloque «{titulo}» reducido a una '
                           f'línea de remisión → {" + ".join(destinos)} '
                           f'({len(filas)} tareas duplicadas fuera del '
                           'contador)')
            hecho = True
            break
    renumerar(ws)


# ==========================================================================
# §2.5 (R3-f) — bandas que duplican el marco: se ANOTAN, no se borran
# ==========================================================================
#: `colapsar_duplicados` reconoce las bandas por su rótulo LITERAL, tomado del
#: representante («CIERRE DE CAJA», «CIERRE GENERAL», «SISTEMAS»). En pizzería
#: las secciones se llaman «SALA», «TERRAZA», «DELIVERY Y CAJA», «BAÑOS» y no
#: colapsaba nada: quedaban duplicadas y CONTADAS mientras las Instrucciones
#: prometían «No se duplican». Aquí se reconocen por lo que de verdad importa
#: —que sus tareas ya estén en 08/09— y se resuelve con una NOTA en la banda:
#: borrar tareas de un hermano sin leerlas una a una es lo que no se puede
#: hacer a ciegas.
NOTA_DUP = ' → estas tareas se detallan en {}'
RX_NOTA_DUP = re.compile(r'→ estas tareas se detallan en ')
UMBRAL_BANDA = 0.8          # ≥80 % de las tareas de la banda ya están en 08/09
UMBRAL_JACCARD = 0.5
UMBRAL_SECUENCIA = 0.75


def _tokens(v):
    t = unicodedata.normalize('NFD', (v or '').lower())
    t = ''.join(ch for ch in t if unicodedata.category(ch) != 'Mn')
    return [w for w in re.split(r'[^a-z0-9]+', t) if len(w) > 2]


def _parecidas(a, b):
    sa, sb = set(a), set(b)
    if not sa or not sb:
        return False
    if len(sa & sb) / len(sa | sb) >= UMBRAL_JACCARD:
        return True
    return difflib.SequenceMatcher(
        None, ' '.join(a), ' '.join(b)).ratio() >= UMBRAL_SECUENCIA


def anotar_duplicados(ws, cambios):
    """Nota «→ estas tareas se detallan en 08/09» en las bandas que duplican."""
    marco = CTX.get('tareas_marco') or {}
    if not marco:
        return
    tok_marco = {f: [_tokens(t) for t in txt] for f, txt in marco.items()}
    g = geometria(ws)
    if not g:
        return
    tope = g['contador'] or ws.max_row
    bandas = [r for r in range(g['hr'] + 1, tope) if es_fila_seccion(ws, r)]
    for i, banda in enumerate(bandas):
        rotulo = ws.cell(row=banda, column=1).value
        if not isinstance(rotulo, str) or RX_NOTA_DUP.search(rotulo):
            continue
        fin = bandas[i + 1] if i + 1 < len(bandas) else tope
        tareas = [forma_estable(ws.cell(row=r, column=2).value)
                  for r in range(banda + 1, fin)
                  if isinstance(ws.cell(row=r, column=1).value, int)
                  and isinstance(ws.cell(row=r, column=2).value, str)
                  and ws.cell(row=r, column=2).value.strip()]
        if len(tareas) < 2:
            continue
        destinos, casadas = [], 0
        for t in tareas:
            tt = _tokens(t)
            de = [f for f, lista in sorted(tok_marco.items())
                  if any(_parecidas(tt, m) for m in lista)]
            if de:
                casadas += 1
                for f in de:
                    if f not in destinos:
                        destinos.append(f)
        ratio = casadas / len(tareas)
        if ratio >= 0.25:
            SOLAPES.append({'hoja': f'{ws.title}!A{banda}',
                            'banda': rotulo.strip(),
                            'tareas': len(tareas), 'casadas': casadas,
                            'ratio': round(ratio, 2),
                            'destinos': sorted(destinos),
                            'anotada': ratio >= UMBRAL_BANDA})
        if ratio < UMBRAL_BANDA or not destinos:
            continue
        ws.cell(row=banda, column=1).value = rotulo.rstrip() + NOTA_DUP.format(
            ' y '.join(sorted(destinos)))
        cambios.append(f'«{ws.title}»: la banda «{rotulo.strip()}» remite a '
                       f'{" y ".join(sorted(destinos))} '
                       f'({casadas}/{len(tareas)} de sus tareas ya están allí)')


def renumerar(ws):
    g = geometria(ws)
    if not g:
        return
    n = 0
    for r in range(g['hr'] + 1, (g['contador'] or ws.max_row)):
        cel = ws.cell(row=r, column=1)
        if isinstance(cel.value, int):
            n += 1
            cel.value = n
    return n


# ==========================================================================
# R3-e — molde P4: UNA sola regla por producto
# ==========================================================================
# Los kits de catering, chocolatería, heladería, hotel y restaurante-creativo
# sólo tienen en alcance su 08/09 (18/19 y 10/11). El motor los dejaba con DOS
# reglas dentro del mismo producto: 08/09 con «✓,—,N/A» y denominador honesto,
# y los otros 9-17 ficheros con «✓,✗,—» y `COUNTIF` sin N/A. El cliente veía
# dos desplegables y dos semánticas de conteo en la misma compra.
#
# Aquí se les aplica lo MÍNIMO que cierra esa brecha —desplegable, contador,
# formato condicional y bio— SIN tocar su maquetación ni sus columnas: no se
# reordena nada, no se insertan filas libres, no se repintan verdes, no se
# reescriben las Instrucciones y no se protege ni se cambia la impresión.
#
# Y de paso corrige un contador que estaba MINTIENDO: el molde P4 repite la
# fila de cabecera en cada sección, así que `COUNTIF(E6:E28,"✓")` contaba los
# «✓» de los ROTULOS y `COUNTIF(B6:B28,"?*")` los «Tarea» de esos mismos
# rótulos. Medido en heladería, 01-apertura-cierre.xlsx:'Apertura'!C29/E29: la
# hoja recién impresa, sin marcar nada, anunciaba «2 de 19» cuando sus tareas
# son 17.
def geometria_p4(ws):
    """Geometría de una hoja del molde P4, o None si no lo es."""
    if geometria(ws):
        return None                     # es del molde ▸: la trata el motor
    hr = cols = None
    for r in range(1, 9):
        if ws.cell(row=r, column=2).value != 'Tarea':
            continue
        if ws.cell(row=r, column=1).value not in ('Nº', '#'):
            continue
        c_ = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                c_[v.strip()] = c
        if CAB_P4 in c_ and CAB_MARCA not in c_:
            hr, cols = r, c_
            break
    if hr is None:
        return None
    contador = None
    for r in range(hr + 1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 4) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and RX_CONTADOR.match(v):
                contador = r
                break
        if contador:
            break
    pie = None
    for r in range(hr + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and (RX_PIE.match(v) or v.startswith('©')
                                   or v.startswith('Firma')):
            pie = r
            break
    nuevo = False
    if contador is None:
        # Las plantillas personalizables del molde P4 se entregan SIN contador
        # ninguno: el cliente escribe sus tareas y no hay nada que le diga
        # cuántas lleva. Se le pone en la fila libre que queda sobre el pie,
        # con el denominador por FÓRMULA (§2.3), sin mover nada de sitio.
        hueco = (pie - 1) if pie else ws.max_row
        if hueco > hr and all(ws.cell(row=hueco, column=c).value is None
                              for c in range(1, ws.max_column + 1)):
            contador, nuevo = hueco, True
        else:
            return None
    return {'hr': hr, 'cols': cols, 'marca': cols[CAB_P4],
            'contador': contador, 'nuevo': nuevo,
            'ncol': ncol_cabecera(ws, hr)}


def _contador_p4(ws, g, cambios):
    m = get_column_letter(g['marca'])
    lo, hi = g['hr'] + 1, g['contador'] - 1
    if hi < lo:
        return
    # La corrección de las cabeceras repetidas se hace por fórmula y no
    # acotando bloques: así el cliente que escriba en una fila en blanco del
    # medio sigue contando, que es lo que hacía el COUNTIF original.
    num = (f'=COUNTIFS(B{lo}:B{hi},"?*",{m}{lo}:{m}{hi},"{MARCA_OK}")'
           f'-COUNTIFS(B{lo}:B{hi},"Tarea",{m}{lo}:{m}{hi},"{MARCA_OK}")')
    den = (f'=COUNTIF(B{lo}:B{hi},"?*")-COUNTIF(B{lo}:B{hi},"Tarea")'
           f'-COUNTIF({m}{lo}:{m}{hi},"N/A")')
    f = g['contador']
    if g['nuevo']:
        cel = ws.cell(row=f, column=2, value=ETIQ_CONTADOR)
        cel.font = Font(bold=True, size=11)
        cambios.append(f'«{ws.title}»: contador con denominador por fórmula '
                       f'(no lo tenía) en la fila {f}')
    antes = (ws.cell(row=f, column=g['marca'] - 2).value,
             ws.cell(row=f, column=g['marca']).value)
    ws.cell(row=f, column=g['marca'] - 2).value = num
    ws.cell(row=f, column=g['marca'] - 1).value = 'de'
    ws.cell(row=f, column=g['marca']).value = den
    reg(ws, ws.cell(row=f, column=g['marca'] - 2).coordinate, num)
    reg(ws, ws.cell(row=f, column=g['marca']).coordinate, den)
    if antes != (num, den) and not g['nuevo']:
        cambios.append(f'«{ws.title}»: contador honesto — el denominador ya no '
                       'cuenta las cabeceras repetidas de cada sección y el '
                       'numerador exige texto en «Tarea»')


def _cf_p4(ws, g):
    ws.conditional_formatting = ConditionalFormattingList()
    lo, hi = g['hr'] + 1, g['contador'] - 1
    if hi < lo:
        return
    _ok = verde_ok()
    ws.conditional_formatting.add(
        f'A{lo}:{get_column_letter(g["ncol"])}{hi}',
        FormulaRule(formula=[f'${get_column_letter(g["marca"])}{lo}'
                             f'="{MARCA_OK}"'],
                    fill=PatternFill('solid', start_color=_ok,
                                     end_color=_ok)))


def _dv_p4(ws, cambios):
    n = 0
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        if not (_es_lista_de_marca(dv.formula1) or dv.formula1 == DV_LISTA):
            continue
        if dv.formula1 != DV_LISTA:
            n += 1
        dv.formula1 = DV_LISTA
        dv.showErrorMessage = True
        dv.errorStyle = 'stop'
        dv.errorTitle = DV_ERROR_TIT
        dv.error = DV_ERROR
    if n:
        cambios.append(f'«{ws.title}»: desplegable {DV_LISTA} (venía con la '
                       'lista vieja «✓,✗,—», que no tiene N/A y deja el '
                       'producto con dos reglas de conteo)')
    return n


def bio_en_instrucciones(wb, cambios):
    """§2.6 sin reescribir la hoja: bio anclada encima de la versión."""
    if 'Instrucciones' not in wb.sheetnames:
        return
    ws = wb['Instrucciones']
    col = 2 if any(isinstance(ws.cell(row=r, column=2).value, str)
                   for r in range(1, min(ws.max_row, 12) + 1)) else 1
    fila_v = tiene_bio = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str):
            continue
        if RX_BIO.search(v):
            tiene_bio = r
        if RX_VERSION.match(v):
            fila_v = r
    if fila_v is None:
        return
    if tiene_bio is None:
        sig = fila_v + 1
        libre = all(ws.cell(row=sig, column=c).value is None
                    for c in range(1, max(ws.max_column, col) + 1))
        if not libre:
            insertar_filas(ws, fila_v)
            fila_v += 1
            sig = fila_v
        estilo = copy.copy(ws.cell(row=fila_v, column=col)._style)
        alto = ws.row_dimensions[fila_v].height
        ws.cell(row=sig, column=col).value = None
        ws.cell(row=sig, column=col)._style = estilo
        ws.row_dimensions[sig].height = alto
        ws.cell(row=fila_v, column=col).value = BIO
        fila_v = sig
        cambios.append('Instrucciones: línea de autoría anclada encima de la '
                       'versión (§2.6)')
    if ws.cell(row=fila_v, column=col).value != version_line():
        ws.cell(row=fila_v, column=col).value = version_line()
        cambios.append('Instrucciones: versión 2.0')


def normalizar_p4(wb, cambios, saltar=(), bio=True):
    """DV, contador honesto, CF y bio en las hojas del molde P4."""
    tocadas, n_dv = [], 0
    for ws in wb.worksheets:
        if ws.title in saltar:
            continue
        n_dv += _dv_p4(ws, cambios)
        if ws.title == 'Instrucciones':
            continue
        g = geometria_p4(ws)
        if not g:
            continue
        _contador_p4(ws, g, cambios)
        _cf_p4(ws, g)
        tocadas.append(ws.title)
    # (j) — la bio y la versión NO dependen de que haya algo que normalizar.
    # Antes se pedían `tocadas or n_dv` y los dos ficheros del kit que no son
    # ni checklist ni P4 —los BONUS-02 de catering y de hotel, calendarios de
    # otro molde— se publicaban con «Versión 1.1» y sin autoría dentro de un
    # producto v2.0: el cliente abre once ficheros y dos le dicen otra versión.
    # Se escribe SÓLO en celdas que ya existen (la línea de versión y la de
    # debajo, vacía): no se toca la maquetación. Si no hay hoja «Instrucciones»
    # —o no hay línea de versión donde anclar— no se inventa nada y lo canta el
    # gate `dv_y_bio` de main.py.
    if bio:
        bio_en_instrucciones(wb, cambios)
    return tocadas


# ==========================================================================
# §1.4 — 08: Responsable y Hora Límite precargados
# ==========================================================================
def precargar_negocio(ws, cambios):
    g = geometria(ws)
    if not g or 'Responsable' not in g['cols']:
        return
    col_r = g['cols']['Responsable']
    col_h = _col_tiempo(g['cols'])
    apertura = ws.title.lower().startswith('apertura')
    filas = [r for r in range(g['hr'] + 1, g['ultima'] + 1)
             if isinstance(ws.cell(row=r, column=1).value, int)]
    if not filas:
        return
    ancla = CTX.get('hora_apertura', '07:00')
    literal = CTX.get('literal_cierre', 'Cierre')
    n = 0
    for i, r in enumerate(filas):
        if apertura:
            resp = 'Encargado'
            hora = _sumar_minutos(ancla, 15 * (i * 3 // max(1, len(filas))))
        else:
            resp = 'Último en salir' if i >= len(filas) - 3 else 'Encargado'
            hora = literal
        if ws.cell(row=r, column=col_r).value != resp:
            ws.cell(row=r, column=col_r).value = resp
            n += 1
        if col_h and ws.cell(row=r, column=col_h).value != hora:
            ws.cell(row=r, column=col_h).value = hora
            n += 1
    if n:
        cambios.append(f'«{ws.title}»: Responsable y Hora Límite precargados '
                       f'en {len(filas)} tareas (ancla {ancla})')


# ==========================================================================
# §1 — CAJA
# ==========================================================================
def caja_columnas(ws, cambios):
    """Columna «Firma» + orden de cabeceras del molde ▸ (TEC-07/COM-21).

    Las hojas de caja venían con «#|Tarea|Responsable|✓|Hora|Notas»: sin firma
    por tarea y con la marca una columna antes de donde está en 01-07, que es
    lo que descolocaba el contador (TEC-21). Queda
    «#|Tarea|Responsable|Hora Límite|✓ Completada|Firma|Notas».
    """
    hr, cols = cabecera_checklist(ws)
    if hr is None:
        return
    # Anchos: escritura ABSOLUTA en cada pasada (fuera del centinela de abajo).
    # TEC-R2-03 — la columna A es la del «#» del checklist y también la de las
    # etiquetas de moneda del recuento: con ancho 5, «0,50 €» (6 caracteres) no
    # cabía en una línea y las cinco monedas se leían igual mientras el cajero
    # contaba calderilla. TEC-R2-07 — «Responsable de caja» (19) en una columna
    # de 16 se leía «Responsable de».
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['C'].width = 20
    if 'Firma' in cols:
        return
    # Centinela: la reordenación es POSICIONAL (inserta en la 6 y reetiqueta
    # de la 4 a la 7), así que sólo puede correr sobre la cabecera exacta que
    # comparten los 12 kits de la familia. Verificado el 2026-08-23 en las 13
    # carpetas kit-tareas* de dl/: las 26 hojas de caja tienen esta misma.
    actual = tuple(ws.cell(row=hr, column=c).value
                   for c in range(1, ws.max_column + 1))
    if actual != CAB_CAJA:
        cambios.append(f'«{ws.title}»: cabecera de caja inesperada {actual} '
                       '— NO se toca (esperada ' + str(CAB_CAJA) + ')')
        return
    insertar_columna(ws, 6)
    ws.cell(row=hr, column=4).value = 'Hora Límite'
    ws.cell(row=hr, column=5).value = CAB_MARCA
    ws.cell(row=hr, column=6).value = 'Firma'
    ws.cell(row=hr, column=7).value = 'Notas'
    for letra, ancho in (('D', 13), ('E', 14), ('F', 14), ('G', 25)):
        ws.column_dimensions[letra].width = ancho
    cambios.append(f'«{ws.title}»: columna «Firma» y orden de cabeceras del '
                   'molde ▸ (Hora Límite · ✓ Completada · Firma · Notas)')


def precargar_caja(ws, cambios):
    g = geometria(ws)
    if not g or 'Responsable' not in g['cols']:
        return
    col_r = g['cols']['Responsable']
    col_h = _col_tiempo(g['cols'])
    apertura = 'apertura' in ws.title.lower()
    ancla = CTX.get('hora_apertura', '07:00')
    literal = CTX.get('literal_cierre', 'Cierre')
    filas = [r for r in range(g['hr'] + 1, g['ultima'] + 1)
             if isinstance(ws.cell(row=r, column=1).value, int)]
    n = 0
    for i, r in enumerate(filas):
        resp = 'Responsable de caja'
        # DOM-R2-08: las 11 tareas llevaban todas la MISMA hora y era ANTERIOR
        # a la apertura del local (06:45 con la alarma puesta y las puertas
        # cerradas: nadie enciende el TPV ni cuenta el fondo entonces). Se
        # anclan DESPUÉS de la apertura y se escalonan como en el 08.
        hora = (_sumar_minutos(ancla, 15 * (1 + i * 3 // max(1, len(filas))))
                if apertura else literal)
        if ws.cell(row=r, column=col_r).value != resp:
            ws.cell(row=r, column=col_r).value = resp
            n += 1
        if col_h and ws.cell(row=r, column=col_h).value != hora:
            ws.cell(row=r, column=col_h).value = hora
            n += 1
    if n:
        cambios.append(f'«{ws.title}»: Responsable y hora precargados')


def moneda_002(ws, cambios):
    """§1.1 / DOM-03 — la moneda de 0,02 € es curso legal y faltaba."""
    r5 = _buscar(ws, '0,05 €', col=1)
    r1 = _buscar(ws, '0,01 €', col=1)
    if r5 is None or r1 is None or _buscar(ws, '0,02 €', col=1):
        return
    insertar_filas(ws, r1)
    for c in range(1, 4):
        ws.cell(row=r1, column=c)._style = copy.copy(
            ws.cell(row=r5, column=c)._style)
    ws.cell(row=r1, column=1).value = '0,02 €'
    ws.cell(row=r1, column=2).value = 0
    f = f'=B{r1}*0.02'
    ws.cell(row=r1, column=3).value = f
    ws.cell(row=r1, column=3).number_format = FMT_EUR
    reg(ws, f'C{r1}', f)
    cambios.append('«Cierre de Caja»: añadida la denominación de 0,02 €')


def recuento(ws, cambios):
    """Cantidades enteras ≥ 0, verdes, y TOTAL EFECTIVO recalculado (TEC-20)."""
    hr = fila_recuento(ws)
    if hr is None:
        return None
    filas = [r for r in range(hr + 1, ws.max_row + 1)
             if isinstance(ws.cell(row=r, column=1).value, str)
             and ws.cell(row=r, column=1).value.strip().endswith('€')]
    if not filas:
        return None
    dv = DataValidation(type='whole', operator='greaterThanOrEqual',
                        formula1='0', allow_blank=True, showErrorMessage=True,
                        errorStyle='stop', errorTitle='Cantidad no válida',
                        error='Escribe el número de billetes o monedas: un '
                              'número entero y positivo, sin decimales.')
    ws.add_data_validation(dv)
    for r in filas:
        cel = ws.cell(row=r, column=2)
        cel.number_format = FMT_ENT
        _verde(cel)
        dv.add(cel)
        ws.cell(row=r, column=3).number_format = FMT_EUR
    # Rayado: la fila de 0,02 € se inserta en medio y rompe la alternancia
    # heredada. Se repinta por tramos contiguos (BILLETES y MONEDAS arrancan
    # los dos en gris), que además es escritura absoluta e idempotente.
    tramo = 0
    for i, r in enumerate(filas):
        if i and r != filas[i - 1] + 1:
            tramo = 0
        color = GRIS if tramo % 2 == 0 else 'FFFFFF'
        for c in (1, 3):
            _relleno(ws.cell(row=r, column=c), color)
        tramo += 1

    total = _buscar(ws, 'TOTAL EFECTIVO', col=1)
    if total:
        f = f'=SUM(C{filas[0]}:C{filas[-1]})'
        ws.cell(row=total, column=3).value = f
        ws.cell(row=total, column=3).number_format = FMT_EUR
        reg(ws, f'C{total}', f)
        ws.row_dimensions[total].height = 28
    # TEC-R2-02 — `insertar_filas` no arrastra las alturas de forma fiable y la
    # fila nueva de 0,02 € (y la cabecera «Denominación») se publicaban a 15 pt
    # entre filas de 22-32. Se fijan SIEMPRE, que además es escritura absoluta.
    ws.row_dimensions[hr].height = 26
    for r in filas:
        ws.row_dimensions[r].height = 22
    cambios.append(f'«{ws.title}»: recuento con validación entera ≥ 0 en '
                   f'{len(filas)} denominaciones y TOTAL EFECTIVO al día')
    return total


def fondo_de_caja(wb, cambios):
    """§1.1 — el fondo de caja NO es una venta (DOM-01/TEC-01/COM-01)."""
    ap = wb['Apertura de Caja']
    fila_fondo = _buscar(ap, ETIQ_FONDO, col=2)
    if fila_fondo is None:
        firma = _buscar_prefijo(ap, 'Firma del responsable:', col=1)
        if firma is None:
            firma = ap.max_row - 1
        insertar_filas(ap, firma, 2)
        fila_fondo = firma
        modelo = firma - 2
        for c in range(1, ap.max_column + 1):
            ap.cell(row=fila_fondo, column=c)._style = copy.copy(
                ap.cell(row=modelo, column=c)._style)
            ap.cell(row=fila_fondo + 1, column=c).value = None
        ap.cell(row=fila_fondo, column=2).value = ETIQ_FONDO
        cel = ap.cell(row=fila_fondo, column=3)
        cel.value = 0
        cel.number_format = FMT_EUR
        _verde(cel)
        ap.cell(row=fila_fondo, column=4).value = (
            'Se descuenta del efectivo en el Cierre de Caja')
        _merge(ap, f'D{fila_fondo}:{get_column_letter(ap.max_column)}'
                   f'{fila_fondo}')
        cambios.append(f'«Apertura de Caja»: celda editable del fondo de caja '
                       f'(C{fila_fondo})')
    else:
        _verde(ap.cell(row=fila_fondo, column=3), FMT_EUR)

    # DOM-05/COM-15 — la tarea apunta a la celda. DOM-R2-18/TEC-R2-04: SIN la
    # coordenada. Iba escrita a mano («…en la celda verde C23»), así que en
    # cuanto el cliente insertase una fila —lo que las propias Instrucciones le
    # invitan a hacer— la fórmula del cierre se reajustaba sola y el texto no.
    # Y con 73 caracteres tampoco cabía en la celda: se recortaba justo por el
    # final, que es donde estaba el dato.
    objetivo = 'Anotar el fondo de caja en la celda verde de abajo'
    for r in range(1, ap.max_row + 1):
        v = ap.cell(row=r, column=2).value
        if not isinstance(v, str):
            continue
        if v == objetivo:
            break
        if 'fondo de caja' in v.lower() and 'registrar' in v.lower():
            ap.cell(row=r, column=2).value = objetivo
            cambios.append('«Apertura de Caja»: la tarea del fondo apunta a la '
                           'celda verde sin hardcodear la coordenada')
            break
    return fila_fondo


def resumen_cierre(wb, fila_fondo, cambios):
    """§1.1 — TOTAL FACTURADO = efectivo − fondo + tarjetas + otros."""
    ci = wb['Cierre de Caja']
    r_ef = _buscar(ci, ETIQ_EFECTIVO, col=1) or _buscar(ci, 'Total Efectivo',
                                                        col=1)
    if r_ef is None:
        return None
    # Dos filas nuevas bajo el recuento: el fondo que se descuenta y —TEC-R2-11
    # / DOM-R2-01— las VENTAS EN EFECTIVO, que es la cifra que el Registro
    # Mensual pide y que el resumen no enseñaba por ninguna parte. Sin ella el
    # operario transcribía el recuento bruto y el registro marcaba un descuadre
    # falso de +fondo todos los días.
    for i, etq in enumerate((ETIQ_FONDO_RESUMEN, ETIQ_VENTAS_EF), start=1):
        if ci.cell(row=r_ef + i, column=1).value != etq:
            insertar_filas(ci, r_ef + i)
            for c in range(1, ci.max_column + 1):
                ci.cell(row=r_ef + i, column=c)._style = copy.copy(
                    ci.cell(row=r_ef, column=c)._style)
            cambios.append(f'«Cierre de Caja»: fila «{etq}» en el Resumen de '
                           'Cierre')
        ci.cell(row=r_ef + i, column=1).value = etq
    ci.cell(row=r_ef, column=1).value = ETIQ_EFECTIVO
    r_fondo_res, r_ventas = r_ef + 1, r_ef + 2

    r_cab = _buscar(ci, 'Concepto', col=1)
    r_tar = _buscar(ci, 'Total Tarjetas (Visa/MC)', col=1)
    r_otr = _buscar(ci, 'Total Otros (Bizum, Vales)', col=1)
    r_tot = _buscar(ci, 'TOTAL FACTURADO', col=1)
    r_z = _buscar(ci, ETIQ_Z, col=1)
    r_desc = _buscar(ci, 'DESCUADRE', col=1)
    if None in (r_tar, r_otr, r_tot, r_z, r_desc):
        return None

    # TEC-03 — la etiqueta va combinada A:B y el importe pasa a C (la columna A
    # mide 5 caracteres: es la del «#» del checklist de arriba)
    filas = [r for r in (r_cab, r_ef, r_fondo_res, r_ventas, r_tar, r_otr,
                         r_tot, r_z, r_desc) if r]
    titulo_res = _buscar_prefijo(ci, '📊', col=1)
    if titulo_res:
        ci.row_dimensions[titulo_res].height = 30
    for r in filas:
        # La etiqueta pasa a ocupar A:B (55 caracteres) y el importe a C, como
        # ya hacía la fila TOTAL EFECTIVO del recuento. Antes la etiqueta vivía
        # en la columna A —ancho 5, la del «#» del checklist de arriba— con
        # wrap y alto fijo: «Total Tarjetas (Visa/MC)» y «Total Otros (Bizum,
        # Vales)» se leían las dos como «Total» (TEC-03).
        # TEC-R2-02 — altura explícita: con la columna A ya en 10 (TEC-R2-03)
        # las etiquetas caben y la tabla del dinero deja de imprimirse
        # aplastada a 15 pt entre filas de 22-32.
        ci.row_dimensions[r] .height = 26 if r == r_cab else 24
        _desmerge_fila(ci, r)
        _merge(ci, f'A{r}:B{r}')
        # El estilo del importe se toma de la ETIQUETA (columna A), que el
        # motor no modifica nunca: así la celda de valor conserva la banda de
        # su fila y la 2.ª pasada llega al mismo sitio. Copiarlo de B no vale:
        # al combinar A:B, B pasa a ser MergedCell y su estilo deriva.
        ci.cell(row=r, column=3)._style = copy.copy(
            ci.cell(row=r, column=1)._style)
        ci.cell(row=r, column=3).alignment = Alignment(horizontal='right',
                                                       vertical='center')
    if r_cab:
        ci.cell(row=r_cab, column=3).value = 'Importe (€)'
        ci.cell(row=r_cab, column=3).alignment = Alignment(
            horizontal='center', vertical='center')

    total_ef = _buscar(ci, 'TOTAL EFECTIVO', col=1)
    formulas = {
        r_ef: f'=C{total_ef}' if total_ef else None,
        r_fondo_res: f"=IFERROR('Apertura de Caja'!C{fila_fondo},0)",
        r_ventas: f'=IFERROR(C{r_ef}-C{r_fondo_res},0)',
        r_tot: f'=C{r_ventas}+C{r_tar}+C{r_otr}',
        r_desc: f'=IFERROR(C{r_tot}-C{r_z},0)',
    }
    for r, f in formulas.items():
        if not f:
            continue
        ci.cell(row=r, column=3).value = f
        ci.cell(row=r, column=3).number_format = FMT_EUR
        reg(ci, f'C{r}', f)
    for r in (r_tar, r_otr, r_z):
        cel = ci.cell(row=r, column=3)
        if not isinstance(cel.value, (int, float)):
            cel.value = 0
        cel.number_format = FMT_EUR
        _verde(cel)

    ci.conditional_formatting.add(f'C{r_desc}', CellIsRule(
        operator='notEqual', formula=['0'],
        fill=PatternFill('solid', start_color=AMBAR, end_color=AMBAR)))
    cambios.append('«Cierre de Caja»: TOTAL FACTURADO = (efectivo − fondo) + '
                   f'tarjetas + otros y DESCUADRE con aviso ámbar (C{r_desc})')
    return r_desc


def registro_mensual(ws, cambios):
    """§1.2 — columna «Z del TPV», descuadre por fórmula y fondo NO sumable."""
    hr = fila_registro_mensual(ws)
    if hr is None:
        return
    cols = {ws.cell(row=hr, column=c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=hr, column=c).value}
    if ETIQ_Z not in cols:
        idx = cols.get('Descuadre')
        if not idx:
            return
        insertar_columna(ws, idx)
        ws.cell(row=hr, column=idx)._style = copy.copy(
            ws.cell(row=hr, column=idx + 1)._style)
        ws.cell(row=hr, column=idx).value = ETIQ_Z
        ws.column_dimensions[get_column_letter(idx)].width = 14
        cambios.append('«Registro Mensual»: columna «Z del TPV» antes de '
                       '«Descuadre» (DOM-04/TEC-04/COM-14)')
        cols = {ws.cell(row=hr, column=c).value: c
                for c in range(1, ws.max_column + 1)
                if ws.cell(row=hr, column=c).value}

    c_fondo = cols.get('Fondo Apertura')
    c_tot = cols.get('Total Facturado')
    c_z = cols[ETIQ_Z]
    c_desc = cols.get('Descuadre')
    filas = [r for r in range(hr + 1, ws.max_row + 1)
             if isinstance(ws.cell(row=r, column=1).value, int)]
    if not filas or not (c_tot and c_desc):
        return

    # DOM-R2-01 — el MISMO libro calculaba el «Total Facturado» de dos maneras
    # incompatibles: el Cierre de Caja descontaba el fondo y el registro no
    # (F=C+D+E), con la columna «Fondo Apertura» de adorno en un AVERAGEIF. Al
    # transcribir el recuento —lo natural, porque la hoja de cierre no enseñaba
    # ninguna otra cifra de efectivo— el registro marcaba +fondo en ámbar TODOS
    # los días. Ahora la columna pide el recuento BRUTO (y lo dice su nombre) y
    # el fondo se descuenta aquí, igual que en el cierre.
    c_ef = None
    for nombre, c in cols.items():
        if isinstance(nombre, str) and (nombre.startswith('Ventas Efectivo')
                                        or nombre.startswith(CAB_EFECTIVO)):
            c_ef = c
    if c_ef and c_fondo and c_ef < c_tot:
        if ws.cell(row=hr, column=c_ef).value != CAB_EFECTIVO:
            ws.cell(row=hr, column=c_ef).value = CAB_EFECTIVO
            cambios.append(f'«{ws.title}»: la columna del efectivo pasa a '
                           f'«{CAB_EFECTIVO}» (el recuento del cajón; el fondo '
                           'lo descuenta la fórmula) — DOM-R2-01')
        lo = get_column_letter(c_ef)
        hi = get_column_letter(c_tot - 1)
        lf = get_column_letter(c_fondo)
        for r in filas:
            # IF(...=0,...) para que un día sin anotar no salga en −fondo.
            f = (f'=IF(SUM({lo}{r}:{hi}{r})=0,0,'
                 f'SUM({lo}{r}:{hi}{r})-{lf}{r})')
            cel = ws.cell(row=r, column=c_tot)
            cel.value = f
            cel.number_format = FMT_EUR
            reg(ws, cel.coordinate, f)

    lz, lt, ld = (get_column_letter(x) for x in (c_z, c_tot, c_desc))
    for r in filas:
        cel = ws.cell(row=r, column=c_z)
        if not isinstance(cel.value, (int, float)):
            cel.value = 0
        cel.number_format = FMT_EUR
        _verde(cel)
        f = f'=IFERROR({lt}{r}-{lz}{r},0)'
        d = ws.cell(row=r, column=c_desc)
        d.value = f
        d.number_format = FMT_EUR
        reg(ws, d.coordinate, f)
        for c in range(2, ws.max_column + 1):
            otra = ws.cell(row=r, column=c)
            if c in (c_tot, c_desc):
                otra.number_format = FMT_EUR
                continue
            if isinstance(otra.value, (int, float)) or otra.value is None:
                if c != cols.get('Responsable'):
                    otra.number_format = FMT_EUR
                _verde(otra)

    totales = None
    for r in range(filas[-1] + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'TOTALES':
            totales = r
            break
    if totales:
        for nombre, c in cols.items():
            if c == 1 or nombre == 'Responsable':
                continue
            letra = get_column_letter(c)
            if c == c_fondo:
                # TEC-05: el fondo es un saldo, no un flujo; sumar 31 fondos no
                # significa nada. AVERAGEIF y no AVERAGE: las 31 filas nacen a 0
                f = f'=IFERROR(AVERAGEIF({letra}{filas[0]}:{letra}{filas[-1]},' \
                    '">0"),0)'
            elif c == c_desc:
                # DOM-R2-26: el descuadre es la única columna que NO se puede
                # sumar con signo — un día de +50 y otro de −50 daban 0 y la
                # fila TOTALES decía que el mes había cuadrado.
                f = (f'=SUMIF({letra}{filas[0]}:{letra}{filas[-1]},">0")'
                     f'-SUMIF({letra}{filas[0]}:{letra}{filas[-1]},"<0")')
            else:
                f = f'=SUM({letra}{filas[0]}:{letra}{filas[-1]})'
            cel = ws.cell(row=totales, column=c)
            cel.value = f
            cel.number_format = FMT_EUR
            reg(ws, cel.coordinate, f)
        nota = ('«' + CAB_EFECTIVO + '» es el recuento del cajón TAL CUAL, con '
                'el fondo dentro: el Total Facturado ya lo descuenta. En '
                'TOTALES, «Fondo Apertura» es el fondo MEDIO de los días '
                'anotados (un saldo no se suma) y «Descuadre» es el descuadre '
                'ACUMULADO en valor absoluto: un +50 y un −50 son 100 € de '
                'descuadre, no 0.')
        rn = totales + 1
        nuevo = ws.cell(row=rn, column=1).value != nota
        ws.cell(row=rn, column=1).value = nota
        ws.cell(row=rn, column=1).font = Font(size=9, italic=True)
        ws.cell(row=rn, column=1).alignment = Alignment(wrap_text=True,
                                                        vertical='top')
        _merge(ws, f'A{rn}:{get_column_letter(ws.max_column)}{rn}')
        ws.row_dimensions[rn].height = None      # que Excel la autoajuste
        if nuevo:
            cambios.append('«Registro Mensual»: nota al pie con las tres '
                           'cifras que NO son una suma (TEC-05/DOM-R2-01/26)')

    ws.conditional_formatting.add(
        f'{ld}{filas[0]}:{ld}{filas[-1]}',
        CellIsRule(operator='notEqual', formula=['0'],
                   fill=PatternFill('solid', start_color=AMBAR,
                                    end_color=AMBAR)))
    cambios.append('«Registro Mensual»: descuadre = IFERROR(Total − Z, 0) en '
                   f'{len(filas)} días con aviso ámbar')


# ==========================================================================
# §2.3 — 07, la plantilla personalizable
# ==========================================================================
#: El aviso «escribe aquí» vive en la BANDA de sección, nunca en la columna
#: «Tarea» (DOM-R2-03/TEC-R2-10/COM-R2-15): como valor de celda, el COUNTIF del
#: denominador lo contaba como tarea y la plantilla en blanco se entregaba
#: marcando «0 de 3» — y quien rellenase una sola sección no llegaba al 100 %
#: nunca. Además, el paréntesis acababa arrastrado a la lista real del cliente.
SUB_TURNO = ('Fecha: ___/___/______    Turno: ☐ Mañana  ☐ Tarde  ☐ Noche    '
             'Responsable turno: _________________________')
SUB_SIMPLE = ('Fecha: ___/___/______    '
              'Responsable: _________________________')
_PLANTILLA_07 = {
    'Por Franja Horaria': {
        'secciones': ('APERTURA', 'SERVICIO', 'CIERRE'),
        'zona': None, 'responsable': None, 'subtitulo': SUB_TURNO,
    },
    'Por Área': {
        'secciones': ('COCINA', 'SALA', 'BARRA'),
        'zona': ('Cocina', 'Sala', 'Barra'), 'responsable': None,
        'subtitulo': SUB_SIMPLE,
    },
    'Por Perfil': {
        'secciones': ('COCINA', 'SALA', 'GERENCIA'),
        'zona': None,
        'responsable': ('Jefe de Cocina', 'Jefe de Sala', 'Gerente'),
        'subtitulo': SUB_SIMPLE,
    },
}
ROTULO_07 = '  {sec} — escribe aquí abajo tus tareas de {min}'
#: Cuando las secciones son PROPIAS del hermano no se repite el nombre en la
#: cola («Apertura (15:00 - 18:00) — … tus tareas de apertura (15:00 -
#: 18:00)»), que además dobla la longitud de la banda.
ROTULO_07_PROPIO = '  {sec} — escribe aquí abajo tus tareas'
#: Las dos colas se quitan para recuperar el NOMBRE de la sección: sin esto la
#: 2.ª pasada leería el rótulo ya escrito por la 1.ª y las secciones propias
#: crecerían una cola en cada ejecución.
RX_SEC_ROTULO = re.compile(
    r'\s+—\s+escribe aquí abajo tus tareas(\s+de\s+.*)?$')
#: Placeholder genérico del generador v1.1 («(Sección 1 — Personaliza este
#: título)» en cafetería, pizzería, hamburguesería y dark-kitchen): eso NO es
#: una sección propia, es la plantilla sin adaptar, y sí se sustituye por el
#: molde del representante.
RX_SEC_GENERICA = re.compile(r'(?i)^\(?\s*secci[óo]n\s*\d')

#: (i) — los hermanos no llaman igual a las tres hojas del 07: bar y pastelería
#: entregan «Por Zona» donde el representante dice «Por Área». Sin el mapa,
#: `diferenciar_07` no reconocía la hoja (se quedaba con «Turno: ☐ Mañana…» en
#: un documento que no es de turno, sin banda de aviso y con la columna «Tarea»
#: sin verde) y la demostración §6 del denominador ni siquiera se ejecutaba,
#: porque exige TRES hojas del 07 en el mismo fichero.
#: «Por Fase» y «Por Turno» (catering, chocolatería, heladería) se quedan FUERA
#: a propósito: son del molde P4, que por diseño recibe sólo el mínimo.
SINONIMOS_07 = {
    'por franja horaria': 'Por Franja Horaria',
    'por franjas horarias': 'Por Franja Horaria',
    'por franja': 'Por Franja Horaria',
    'por franjas': 'Por Franja Horaria',
    'por area': 'Por Área',
    'por areas': 'Por Área',
    'por zona': 'Por Área',
    'por zonas': 'Por Área',
    'por perfil': 'Por Perfil',
    'por perfiles': 'Por Perfil',
}


def _sin_tildes(v):
    t = unicodedata.normalize('NFD', (v or '').strip().lower())
    return ''.join(ch for ch in t if unicodedata.category(ch) != 'Mn')


class _Plantillas07(dict):
    """`dict` que además entiende los sinónimos de título de §2.3.

    Se resuelve aquí y no en cada llamada porque `PLANTILLA_07` se consulta
    con `in` desde tres sitios distintos (`contexto`, los ritmos y la
    demostración §6 de `main.py`): con un mapa aparte, cualquiera de ellos se
    quedaría con la lista corta sin que nada lo cantase.
    """

    def canonica(self, titulo):
        if not isinstance(titulo, str):
            return None
        if dict.__contains__(self, titulo):
            return titulo
        return SINONIMOS_07.get(_sin_tildes(titulo))

    def __contains__(self, titulo):
        return self.canonica(titulo) is not None

    def get(self, titulo, defecto=None):
        clave = self.canonica(titulo)
        return dict.get(self, clave, defecto) if clave else defecto


PLANTILLA_07 = _Plantillas07(_PLANTILLA_07)


def _nombre_seccion(v):
    """Nombre de una banda del 07, sin la cola que escribe el motor."""
    return RX_SEC_ROTULO.sub('', (v or '').strip()).strip()


def _zona_de_seccion(ws, g, banda, fin, col_zona):
    """Primer valor de «Zona» ya escrito en esa sección, o None."""
    for r in range(banda + 1, fin):
        if not isinstance(ws.cell(row=r, column=1).value, int):
            continue
        v = ws.cell(row=r, column=col_zona).value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def cfg_07(ws, g, bandas):
    """Config de §2.3 para esta hoja: la del representante, o la SUYA.

    (i). El molde del representante sólo se impone cuando la hoja no tiene
    secciones propias que perder: o son el placeholder «(Sección N —
    Personaliza este título)» del generador v1.1, o ya son las canónicas. Bar
    entrega CUATRO secciones escritas para un bar («Barra Principal»,
    «Coctelería», «Bodega / Almacén», «Terraza» / «Head Bartender»,
    «Bartender», «Barback», «Camarero») y su Zona precargada: escribirle
    encima COCINA / SALA / BARRA sería borrar el contenido del hermano para
    imponerle el del kit base. Lo que se aplica en ese caso es lo UNIVERSAL —
    subtítulo sin «Turno:» donde no hay turno, banda con el aviso de dónde se
    escribe, columna «Tarea» vacía y en verde, responsable precargado— con sus
    propios nombres.
    """
    cfg = PLANTILLA_07.get(ws.title)
    if not cfg:
        return None
    nombres = [_nombre_seccion(ws.cell(row=b, column=1).value) for b in bandas]
    canon = list(cfg['secciones'])
    if len(bandas) == len(canon) and (
            all(RX_SEC_GENERICA.match(n) for n in nombres)
            or [n.upper() for n in nombres] == [c.upper() for c in canon]):
        return dict(cfg, propias=False)
    if not all(nombres):
        return None                      # bandas sin rótulo: no se inventa
    col_zona = g['cols'].get('Zona')
    zona = None
    if cfg['zona'] and col_zona:
        tope = g['contador'] or ws.max_row
        zona = []
        for i, b in enumerate(bandas):
            fin = bandas[i + 1] if i + 1 < len(bandas) else tope
            zona.append(_zona_de_seccion(ws, g, b, fin, col_zona)
                        or nombres[i])
        zona = tuple(zona)
    return {'secciones': tuple(nombres), 'zona': zona,
            'responsable': tuple(nombres) if cfg['responsable'] else None,
            'subtitulo': cfg['subtitulo'], 'propias': True}


def diferenciar_07(ws, cambios):
    """Las tres hojas eran la misma renombrada (DOM-07/TEC-13/COM-16)."""
    g = geometria(ws)
    if not g or ws.title not in PLANTILLA_07:
        return
    col_zona = g['cols'].get('Zona')
    col_resp = g['cols'].get('Responsable')
    bandas = [r for r in range(g['hr'] + 1, (g['contador'] or ws.max_row))
              if es_fila_seccion(ws, r)]
    cfg = cfg_07(ws, g, bandas) if bandas else None
    if not cfg:
        return
    plantilla = ROTULO_07_PROPIO if cfg['propias'] else ROTULO_07
    n = 0
    # COM-R2-14 — «Turno: ☐ Mañana ☐ Tarde ☐ Noche» en «Por Área» y «Por
    # Perfil», que no son documentos de turno: el mismo «plantilla sin adaptar»
    # que se corrigió en 03, y en el fichero que el cliente hace suyo.
    if ws.cell(row=2, column=1).value != cfg['subtitulo']:
        ws.cell(row=2, column=1).value = cfg['subtitulo']
        n += 1
    for i, banda in enumerate(bandas):
        sec = cfg['secciones'][i]
        rotulo = plantilla.format(sec=sec, min=sec.lower())
        if ws.cell(row=banda, column=1).value != rotulo:
            ws.cell(row=banda, column=1).value = rotulo
            n += 1
        fin = bandas[i + 1] if i + 1 < len(bandas) else (g['contador']
                                                         or ws.max_row)
        primera = True
        for r in range(banda + 1, fin):
            if not isinstance(ws.cell(row=r, column=1).value, int):
                continue
            if col_zona:
                z = cfg['zona'][i] if cfg['zona'] else None
                if ws.cell(row=r, column=col_zona).value != z:
                    ws.cell(row=r, column=col_zona).value = z
                    n += 1
            if col_resp:
                rp = cfg['responsable'][i] if cfg['responsable'] else \
                    ('(Responsable)' if primera else None)
                if ws.cell(row=r, column=col_resp).value != rp:
                    ws.cell(row=r, column=col_resp).value = rp
                    n += 1
            if ws.cell(row=r, column=2).value is not None:
                ws.cell(row=r, column=2).value = None
                n += 1
            # DOM-R2-15 — en el único fichero cuyo propósito es que el cliente
            # escriba, la columna «Tarea» era la ÚNICA que no iba en verde: el
            # código de color señalaba las 12 casillas donde no se escribe la
            # tarea y dejaba en blanco las 12 donde sí.
            _verde(ws.cell(row=r, column=2))
            primera = False
    if n:
        if cfg['propias']:
            cambios.append(
                f'«{ws.title}»: se CONSERVAN sus {len(cfg["secciones"])} '
                'secciones propias ('
                + ' · '.join(cfg['secciones']) + ') y se le aplica el molde ▸ '
                '(subtítulo sin «Turno» donde no hay turno, banda con el aviso '
                f'de dónde se escribe, «Tarea» vacía y en verde) — {n} celdas')
        else:
            cambios.append(f'«{ws.title}»: secciones '
                           f'{" / ".join(cfg["secciones"])} y columnas propias '
                           f'({n} celdas) — las tres hojas dejan de ser la '
                           'misma')


# ==========================================================================
# §2.8 — BONUS
# ==========================================================================
def calendario(ws, cambios):
    hr = fila_calendario(ws)
    if hr is None:
        return None
    ncol = ncol_cabecera(ws, hr)
    libres = [r for r in range(hr + 1, ws.max_row + 1)
              if ws.cell(row=r, column=1).value == '(Tu fecha)']
    n = 0
    for r in libres:
        cel = ws.cell(row=r, column=4)
        if cel.value != '(Antelación)':
            cel.value = '(Antelación)'
            n += 1
        cel._style = copy.copy(ws.cell(row=r, column=3)._style)
        _verde(cel)
    if n:
        cambios.append(f'«{ws.title}»: placeholder y verde en la columna '
                       f'«Antelación» de las {len(libres)} filas libres '
                       '(TEC-24)')
    # DOM-R2-20 — al insertar las fechas nuevas el rayado quedó con dos pares
    # de filas grises consecutivas. En una tabla de 22 filas impresa en
    # apaisado, el zebra es lo único que guía el ojo. Se repinta por ÍNDICE,
    # que además es escritura absoluta.
    eventos = [r for r in range(hr + 1, ws.max_row + 1)
               if r not in libres
               and isinstance(ws.cell(row=r, column=2).value, str)
               and ws.cell(row=r, column=2).value.strip()
               and not es_fila_seccion(ws, r)]
    for i, r in enumerate(eventos):
        color = GRIS if i % 2 == 0 else 'FFFFFF'
        # El ancho sale de la CABECERA: el calendario de bar tiene 6 columnas
        # («# | Fecha | Evento | Preparación Especial | Antelación | Notas») y
        # con las 4 fijas de antes el rayado se cortaba a media tabla.
        for c in range(1, ncol + 1):
            _relleno(ws.cell(row=r, column=c), color)
    if eventos:
        cambios.append(f'«{ws.title}»: rayado rehecho por índice en las '
                       f'{len(eventos)} fechas (DOM-R2-20)')
    return (hr + 1, max(eventos + libres)) if (eventos or libres) else None


def briefing(ws, cambios):
    """DOM-R2-16 — el formulario del briefing no marcaba dónde se escribe.

    20 campos para rellenar y una sola celda en verde, en el fichero que la
    landing manda «imprimir y pegar en el pase». El resto del kit enseña que
    verde = celda de entrada; aquí el código de color no se aplicaba.
    """
    if not es_briefing(ws):
        return []
    campos = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.rstrip().endswith(':'):
            _verde(ws.cell(row=r, column=3))
            campos.append(r)
    if campos:
        cambios.append(f'«{ws.title}»: {len(campos)} casillas de respuesta en '
                       'verde (DOM-R2-16)')
    return campos


# ==========================================================================
# §2.5 / §2.6 — Instrucciones
# ==========================================================================
F_TITULO = Font(bold=True, size=18, color='00FFD700')
F_CAB = Font(bold=True, size=12, color='00333333')
F_TXT = Font(size=11, color='00555555')
ANCHO_B = 100

#: CB-E8 — «Dónde encaja este fichero» es el encabezado que sustituye a «Se
#: conecta con» en el ÚNICO caso en que el bloque no conecta con nadie (el
#: fichero de áreas de un kit sin negocio ni caja). Tiene que estar en esta
#: lista o la 2.ª pasada no lo reconocería como bloque del motor, lo conservaría
#: como texto heredado y volvería a emitir el suyo debajo: el gate de
#: idempotencia de `main.py` lo vería como una diferencia.
MIS_BLOQUES = ('Cómo cuenta el contador', 'Filas libres', 'Protección de la '
               'hoja', 'Se conecta con', 'Dónde encaja este fichero',
               'Qué resuelve', 'Cómo usar',
               'Celdas editables', 'Registro Mensual')
#: CB-E4 — «Cómo personalizar» NO entra en `MIS_BLOQUES`, y la diferencia es de
#: 46 celdas: esa tupla se descarta SIEMPRE, y `kit-tareas/01-apertura-cierre
#: .xlsx` —el representante, publicado— trae ese mismo encabezado con sus dos
#: viñetas. Metiéndolo ahí, el dry-run del representante lo BORRABA sin poner
#: nada en su lugar (el bloque nuevo sólo se emite en la sub-familia CB). Se
#: descarta por `RX_PERSONALIZAR` y sólo cuando `sub_cb()` va a reemitirlo.

#: CB-E4 — el encabezado LITERAL que traen los 77 ficheros de la sub-familia
#: («Como personalizar», sin tilde) y que `ortografia()` acentuará. Se compara
#: sin tildes para que dé igual el orden de los pasos. Los 13 kits de la
#: familia NO lo tienen —cafetería escribe «Consejos de personalización» y
#: hotel «Personalización:»—, así que la sustitución no los alcanza: es la
#: condición que mantiene la regresión en 0.
RX_PERSONALIZAR = re.compile(r'(?i)^c[oó]mo personalizar$')


def _bloque_personalizar(papel):
    """CB-E4 — «Cómo personalizar», escrito para el TIPO de entregable.

    El bloque heredado es un literal único, copiado en los 11 ficheros de cada
    kit, y en tres de ellos describe columnas que no existen. Medido en el
    representante: `03-seguridad-anisakis-appcc.xlsx:Instrucciones!B14/B17`
    habla de «responsables y horarios» y de «horas límite» dentro de un libro
    de REGISTROS (que no tiene ni responsables ni horas), y
    `BONUS-01-briefing-servicio.xlsx:Instrucciones!B13` promete «celdas verdes»
    en una hoja con CERO rellenos. Cierra TEC-10 y COM-19.
    """
    comun = ('Nada lleva contraseña: si necesitas cambiar la estructura, '
             'Revisar → Desproteger hoja.')
    if papel == 'registro':
        return [
            ('h', 'Cómo personalizar'),
            ('b', 'Esto es un REGISTRO, no un checklist: cada fila es un lote, '
                  'un equipo o un día, y se rellena a mano el día que toca. No '
                  'hay responsables ni horas límite que ajustar.'),
            ('b', 'Las columnas de fecha ya vienen con formato DD/MM/AAAA y '
                  'las de temperatura en «0,0 °C»: escribe sólo el número. Si '
                  'una lectura se sale del rango que declara su propia fila, '
                  'la celda se pone en ROJO sola.'),
            # El motor NO afirma nada sobre la norma: el encabezado de estas
            # hojas cita «RD 1420/2006», que el research de guías del
            # 2026-08-29 (commit 837ef05) documenta DEROGADO por el
            # RD 1021/2022 art. 8.1. Quién cita qué es trabajo del módulo de
            # contenido (§2.0, lista blanca normativa); aquí el texto se
            # limita a explicar cómo se rellena la hoja.
            ('b', 'Añade equipos o especies escribiendo en las filas verdes '
                  'libres. Cambiar una fila no cambia el límite legal que le '
                  'aplique: si tu proceso es distinto, revísalo con tu plan '
                  'APPCC antes de tocar el valor.'),
            ('b', comun),
        ]
    if papel == 'formulario':
        return [
            ('h', 'Cómo personalizar'),
            ('b', 'Esto es un FORMULARIO para imprimir y rellenar a mano, no '
                  'una tabla de tareas: no tiene celdas verdes ni desplegable '
                  'porque se escribe sobre el papel.'),
            ('b', 'Cambia los rótulos que no encajen con tu servicio y borra '
                  'los apartados que no uses; la hoja está pensada para caber '
                  'en un A4 y repartirse en el pase.'),
            ('b', comun),
        ]
    if papel == 'calendario':
        return [
            ('h', 'Cómo personalizar'),
            ('b', 'Esto es un CALENDARIO anual: las fechas y las campañas son '
                  'las del calendario español. Cambia las de tu zona (fiestas '
                  'locales, ferias, temporada alta) escribiendo encima.'),
            ('b', 'La antelación es orientativa y está pensada para que la '
                  'compra y la carta lleguen a tiempo: ajústala a tus '
                  'proveedores.'),
            ('b', comun),
        ]
    return [
        ('h', 'Cómo personalizar'),
        ('b', 'Escribe encima de las celdas VERDES: Responsable, la columna de '
              'tiempo, la marca y la firma. El texto de la tarea también es '
              'tuyo — cámbialo a tu vocabulario.'),
        ('b', 'Lo que no aplique en tu local se marca «N/A» en el desplegable, '
              'no se borra: así sale del total y el porcentaje sigue siendo '
              'verdad.'),
        ('b', comun),
    ]


def papel_instrucciones(recon):
    """CB-E4 — tipo de entregable a partir de las hojas RECONOCIDAS."""
    tipos = set(recon.values())
    if 'checklist' in tipos:
        return 'checklist'
    if 'registro_appcc' in tipos or 'registro' in tipos:
        return 'registro'
    if 'briefing' in tipos:
        return 'formulario'
    if 'calendario' in tipos:
        return 'calendario'
    return 'checklist'

#: DOM-R2-30/COM-R2-09 — líneas heredadas de la v1.1 que dicen lo CONTRARIO que
#: los bloques nuevos, cuatro líneas más arriba y por tanto leídas antes: «borra
#: lo que no aplique» (ahora se marca N/A) y «añade en las filas vacías del
#: final» (ahora hay 5 filas verdes DENTRO del rango, y escribir por debajo de
#: la última es justamente lo que no cuenta nadie).
RX_OBSOLETO = re.compile(
    r'(?i)(borra las tareas que no aplican'
    r'|a[ñn]ade tareas espec[ií]ficas'
    r'|filas vac[ií]as del final'
    r'|hojas en blanco)')


def _bloques_contador():
    return [
        ('h', 'Cómo cuenta el contador'),
        ('b', 'Marca con ✓ en la columna «✓ Completada» (desplegable): es la '
              'que cuenta el total de tareas completadas.'),
        # DOM-R2-02: «—» NO sale del total. Antes salía, y las Instrucciones lo
        # vendían como una ventaja: el turno que se saltaba tareas las marcaba
        # «—» y la hoja que el kit manda archivar como prueba imprimía 100 %.
        ('b', '«N/A» = no aplica en tu local: esa tarea SALE del total (no la '
              'borres, márcala N/A). «—» = no hecha: sigue contando como '
              'pendiente y baja el porcentaje, que es de lo que se trata.'),
        ('b', 'El denominador cuenta las tareas escritas en la columna '
              '«Tarea», no un número fijo: si añades o borras tareas, el '
              'total se ajusta solo.'),
        ('h', 'Filas libres'),
        ('b', f'Al final de cada tabla hay {HOLGURA} filas verdes libres '
              'DENTRO del rango que cuenta el contador: escribe ahí tus '
              'tareas propias.'),
        ('b', 'Si necesitas más, inserta filas DENTRO de la tabla (clic '
              'derecho → Insertar), nunca por debajo de la última: lo que se '
              'escriba fuera del rango no lo cuenta nadie.'),
    ]


def _bloque_plantilla():
    """COM-R2-11 — las tres hojas del 07 ya no son «hojas en blanco»."""
    return [
        ('h', 'Qué resuelve'),
        ('b', 'Tres plantillas maestras ya estructuradas —por franja horaria, '
              'por área y por perfil— con las secciones puestas y la zona o el '
              'responsable precargados: tú sólo escribes tus tareas.'),
        ('b', 'Escribe en las celdas VERDES de la columna «Tarea», bajo la '
              'sección que corresponda. El contador arranca en 0 y sube a '
              'medida que escribes: cuenta lo que hay, no un número fijo.'),
    ]


def _bloque_proteccion():
    return [
        ('h', 'Protección de la hoja'),
        ('b', 'Las hojas con fórmulas van protegidas SIN contraseña: las '
              'celdas de entrada (las verdes) están desbloqueadas y los '
              'cálculos no se pisan por accidente.'),
        ('b', 'Puedes insertar y borrar filas con la protección puesta. Para '
              'cambiar la estructura: Revisar → Desproteger hoja.'),
    ]


def _bloque_conecta(fname):
    neg, caja, areas = (CTX.get('f_negocio'), CTX.get('f_caja'),
                        CTX.get('f_areas'))
    # CB-E8 — ESPEJO: el fichero de áreas de un kit que no tiene ni fichero de
    # negocio ni fichero de caja. Es el mismo defecto que T-01 (una rama que no
    # comprueba el caso que la contradice) en la única ranura que quedaba sin
    # comprobar. Medido el 2026-08-29 en el dry-run de sushi-bar:
    # `01-apertura-cierre-sushi.xlsx:Instrucciones!B35` imprimía
    # «▸ 01-apertura-cierre-sushi.xlsx — el mismo día con el DETALLE por área
    # (barra sushi).» DENTRO del propio 01, y B36 remataba con «▸ Estás en
    # 01-apertura-cierre-sushi.xlsx.» (la cola genérica se queda sin marco que
    # citar y degenera). Un bloque titulado «Se conecta con» cuya única entrada
    # es el fichero que el cliente tiene abierto no es un mapa del kit: es un
    # espejo.
    #
    # El caso se acota a `fname == areas and not neg and not caja` a propósito,
    # y no es una restricción cosmética: en los kits con 08/09 la enumeración
    # SÍ es un mapa de los tres niveles y el fichero de áreas aparece en ella
    # por derecho propio —igual que el de negocio aparece en la suya desde
    # T-01—, así que tocar ese caso cambiaría los 11 kits publicados. Medido
    # sobre `dl/` el 2026-08-29: los únicos productos con `f_areas` y sin
    # `f_negocio`/`f_caja` son sushi-bar y asador, ninguno de ellos publicado
    # en v2.0. Gate de regresión: `regresion.py` sobre cafetería y hotel, 0
    # diferencias.
    espejo = bool(areas) and fname == areas and not neg and not caja
    lineas = [('h', 'Dónde encaja este fichero' if espejo
               else 'Se conecta con')]
    orden = []
    # T-04 — los paréntesis salen de la ESTRUCTURA del kit (bandas reales del
    # negocio, nombres reales de hoja del fichero de áreas). Antes iban
    # hardcodeados para los 12 kits y dark-kitchen imprimía «(accesos, luces,
    # clima, terraza)» y «(cocina, sala, barra)» en sus 11 ficheros teniendo
    # sólo «Apertura Cocina» y «Cierre Cocina». Sin datos → sin paréntesis.
    if neg:
        lineas.append(('b', f'{neg} — checklist del LOCAL completo: es el '
                            'MARCO del día'
                            + parentesis(CTX.get('negocio_bandas')) + '.'))
        orden.append('local')
    if areas and not espejo:
        lineas.append(('b', f'{areas} — el mismo día con el DETALLE por área'
                            + parentesis(CTX.get('areas_nombres')) + '.'))
        orden.append('áreas')
    if caja:
        # m6 — el fichero del dinero se describe con el modelo que TIENE. En
        # catering no hay cajón que arquear: hay anticipos, factura y saldo.
        if es_modelo_eventos():
            lineas.append(('b', f'{caja} — la FACTURACIÓN: anticipos, '
                                'liquidación y cobro de cada evento.'))
            orden.append('eventos')
        else:
            lineas.append(('b', f'{caja} — la CAJA: fondo, recuento por '
                                'denominaciones, Z del TPV y descuadre.'))
            orden.append('caja')
    if len(orden) > 1:
        # T-08 — la promesa impresa era más fuerte que lo que el motor
        # garantiza (el umbral del 80 % deja pasar solapes del 25-40 %).
        # m1 — y la frase enumera SÓLO los niveles que este kit tiene: sin
        # fichero de áreas, «el de áreas detalla CÓMO…» contradecía al «Orden
        # de uso: local → caja» de la misma viñeta.
        lineas.append(('b', 'Orden de uso: ' + ' → '.join(orden) + '. '
                            + frase_niveles()))
    # DOM-R2-17/COM-R2-13 — hasta aquí el bloque es el MISMO en los 11 ficheros
    # y sólo nombra tres. El lector de las semanales acababa leyendo «local →
    # áreas → caja. Estás en 05-…», que no es ninguna de las tres cosas, y el
    # calendario anual no remitía al fichero donde están las tareas de San
    # Valentín y Navidad que él mismo anuncia. La última línea se personaliza.
    cal, brief = CTX.get('f_calendario'), CTX.get('f_briefing')
    plant, evt, per = (CTX.get('f_plantilla'), CTX.get('f_eventos'),
                       CTX.get('f_periodico'))
    cola = None
    # T-01 — el fichero que ES el marco no puede remitir a OTRO como marco. En
    # producción, dl/kit-tareas/08-apertura-cierre-negocio.xlsx:Instrucciones
    # !B36 decía «08 … es el MARCO del día» y cuatro líneas más abajo !B40
    # «Estás en 08 …: el marco del día está en 01-apertura-cierre.xlsx»: dos
    # afirmaciones opuestas en la misma hoja impresa. La rama genérica excluía
    # `fname` de los candidatos pero nunca comprobaba si `fname` ERA el marco.
    if neg and fname == neg:
        detalle = []
        if areas:
            detalle.append('el DETALLE de cada zona está en ' + areas)
        if caja:
            detalle.append('el DINERO, en ' + caja)
        cola = ('Estás en ' + fname + ': este ES el marco del día — el HITO de '
                'abrir y cerrar el local'
                + ('; ' + ' y '.join(detalle) if detalle else '') + '.')
    elif caja and fname == caja:
        otros = []
        if neg:
            otros.append('el marco del día está en ' + neg)
        if areas:
            otros.append('el detalle por zona, en ' + areas)
        if es_modelo_eventos():
            # m6 — (e) FIRMADO se escribió para el modelo de mostrador. Aquí el
            # mismo hueco lo ocupa la facturación por evento: mantener el
            # literal de caja diría «fondo, recuento, Z del TPV y descuadre» en
            # el único fichero del kit donde no hay ninguna de las cuatro.
            cola = ('Estás en ' + fname + ': esta es la FACTURACIÓN — el '
                    'DINERO de cada evento (anticipo, liquidación, factura y '
                    'saldo pendiente)'
                    + ('; ' + ' y '.join(otros) if otros else '') + '.')
        else:
            cola = ('Estás en ' + fname + ': esta es la CAJA — el DINERO del '
                    'día (fondo, recuento, Z del TPV y descuadre)'
                    + ('; ' + ' y '.join(otros) if otros else '') + '.')
    elif espejo:
        # CB-E8 — la cola que le faltaba a la ranura de ÁREAS, simétrica a la
        # del negocio y a la de la caja. Antes caía en la genérica, que compone
        # el marco con `(neg, areas)` excluyendo `fname`: sin negocio y siendo
        # uno mismo las áreas, el marco salía vacío y la frase degeneraba en
        # «Estás en 01-….», que no dice absolutamente nada.
        cola = ('Estás en ' + fname + ': este ES el DETALLE por área del día'
                + parentesis(CTX.get('areas_nombres'))
                + ' — cómo se abre y se cierra cada zona.')
        if CTX.get('sin_caja'):
            # Sólo con CB-E7 puesto se puede AFIRMAR que no hay fichero del
            # dinero en ningún sitio: `espejo` mira las dos ranuras, y
            # `sin_caja` mira además las cuatro firmas del dinero en los 11
            # ficheros, estén o no en alcance.
            cola += (' Este kit no trae un fichero aparte de negocio ni de '
                     'caja: el marco del día es este mismo.')
    elif fname == cal and evt:
        cola = (f'Estás en {fname}: cada fecha de este calendario se ejecuta '
                f'con los checklists de {evt}.')
    elif fname == brief:
        destino = areas or neg
        cola = (f'Estás en {fname}: el briefing abre el turno que luego se '
                f'trabaja con {destino}' + (f' y {evt}' if evt else '') + '.')
    elif fname == plant:
        cola = (f'Estás en {fname}: aquí construyes TUS listas cuando ninguna '
                'de las otras encaja. Copia la estructura de la que más se '
                'parezca.')
    elif fname == evt:
        cola = (f'Estás en {fname}: lo excepcional (eventos, festivos, '
                'temporada). El día a día está en '
                + ' y '.join(x for x in (neg, areas) if x) + '.')
    elif fname == per:
        cola = (f'Estás en {fname}: lo que NO es diario (semanal, mensual, '
                'trimestral y anual). El día a día está en '
                + ' y '.join(x for x in (neg, areas) if x) + '.')
    if not cola:
        # Ni siquiera el caso genérico se queda en «Estás en X.»: el lector
        # tiene que saber dónde encaja su hoja dentro del kit.
        marco = ' y '.join(x for x in (neg, areas) if x and x != fname)
        cola = (f'Estás en {fname}: es una capa MÁS, no una repetición — el '
                f'marco del día está en {marco}.') if marco else \
               ('Estás en ' + fname + '.')
    lineas.append(('b', cola))
    # El bloque de dos líneas (encabezado + cola) se descarta por inútil: es el
    # caso de un kit sin ninguna ranura de papel, donde la cola genérica se
    # queda en «Estás en X.». CB-E8 es la excepción: ahí las dos líneas son un
    # encabezado honesto y una cola que dice qué es el fichero.
    return lineas if (len(lineas) > 2 or espejo) else []


def instrucciones_caja(fname):
    kit = CTX.get('kit', '')
    titulo = 'Apertura y Cierre de Caja' + (f' — {kit}' if kit else '')
    return titulo, [
        ('h', 'Qué resuelve'),
        ('b', 'El arqueo diario: fondo de caja, recuento del cajón por '
              'denominaciones, Z del TPV y descuadre — y el registro mensual '
              'para ver de un vistazo qué días no cuadraron.'),
        ('h', 'Cómo usar — Apertura de Caja'),
        ('b', 'Cuenta el fondo y escribe el importe en la celda VERDE «Fondo '
              'de caja inicial (€)». De ahí lo lee el cierre: no hay que '
              'volver a teclearlo.'),
        ('h', 'Cómo usar — Cierre de Caja'),
        ('b', 'En «Recuento de Efectivo por Denominación» rellena SOLO la '
              'columna «Cantidad» (número de billetes o monedas, entero): el '
              'subtotal y el TOTAL EFECTIVO se calculan solos.'),
        ('b', 'En el «Resumen de Cierre» sólo escribes tres cifras: Total '
              'Tarjetas, Total Otros y la Z del TPV. Lo demás es fórmula.'),
        ('b', 'El fondo se descuenta solo: TOTAL FACTURADO = (efectivo '
              'contado − fondo) + tarjetas + otros.'),
        ('b', 'DESCUADRE = TOTAL FACTURADO − Z del TPV. Si la casilla se pone '
              'ÁMBAR, la caja no cuadra: revísalo ANTES de firmar.'),
        ('h', 'Registro Mensual'),
        ('b', 'Una fila por día: escribe fondo, ventas por medio de pago, Z '
              'del TPV y depósito; el Total Facturado y el Descuadre se '
              'calculan solos y el descuadre se pone ámbar.'),
        ('b', 'El fondo de caja es un saldo, no una venta: la fila TOTALES '
              'muestra el fondo MEDIO de los días anotados, no su suma.'),
        ('h', 'Celdas editables'),
        ('b', 'Todo lo que se rellena a diario va en VERDE: Responsable, Hora '
              'Límite, ✓ Completada, Firma, Notas, las cantidades del '
              'recuento y los importes del resumen.'),
    ]


def instrucciones_cobros(fname):
    """m6 — Instrucciones del fichero del dinero del modelo POR EVENTOS.

    `instrucciones_caja` describe un arqueo de mostrador de punta a punta
    (fondo, recuento por denominaciones, Z del TPV, descuadre, registro
    mensual). Aplicárselo al 09 de catering publicaría un manual de uso de
    cuatro hojas que ese fichero no tiene: el cliente leería cómo contar
    calderilla en un producto que factura por transferencia. El vocabulario es
    el de catering —evento, comensales, anticipo, factura, saldo—; «caja» sólo
    aparece en la sección OPCIONAL de la barra en efectivo, que es la única
    parte donde hay un cajón.
    """
    kit = CTX.get('kit', '')
    titulo = 'Cobros y Facturación por Evento' + (f' — {kit}' if kit else '')
    return titulo, [
        ('h', 'Qué resuelve'),
        ('b', 'El dinero de cada EVENTO de principio a fin: qué hay que tener '
              'cerrado antes (presupuesto firmado, anticipo cobrado, datos de '
              'facturación), qué hay que hacer después (comensales reales, '
              'extras, factura, saldo) y cuánto queda por cobrar.'),
        ('b', 'No es un arqueo de caja: una empresa de catering no tiene '
              'mostrador ni turno de TPV. Cobra por evento y casi siempre por '
              'transferencia, así que lo que hay que vigilar es el SALDO '
              'pendiente y su vencimiento, no el descuadre del cajón.'),
        ('h', 'Cómo usar — Antes del Evento'),
        ('b', 'Repásalo con el calendario en la mano: la columna «Cuándo» va '
              'en días ANTES del evento (D-15, D-7, D-3, D-1), no en horas. '
              'Nada sale de aquí sin presupuesto aceptado por escrito y sin el '
              'anticipo cobrado y registrado.'),
        ('h', 'Cómo usar — Después del Evento'),
        ('b', 'Los días siguientes (D+0, D+1, D+7, D+30): comensales reales '
              'frente a contratados, extras consumidos, factura emitida, saldo '
              'comunicado, cobro registrado y conciliación con el banco.'),
        ('h', 'Cómo usar — Liquidación del Evento'),
        ('b', 'Escribe sólo las celdas VERDES: presupuesto y extras (base, sin '
              'IVA), cómo se reparte esa base entre el 10 % y el 21 %, el '
              'anticipo ya cobrado, lo cobrado tras el evento y la fecha de '
              'vencimiento del saldo. El total, los IVA, el saldo, el '
              'pendiente y el ESTADO son fórmula.'),
        ('b', 'El reparto del IVA lo decides tú con tu asesor: la comida y las '
              'bebidas no alcohólicas del servicio de catering suelen ir al '
              '10 % y los alquileres, la decoración, los servicios y las '
              'bebidas alcohólicas al 21 %. Si las dos bases no suman '
              'presupuesto + extras, la hoja te avisa.'),
        ('b', 'PENDIENTE DE COBRO se pone en ÁMBAR mientras quede saldo, y el '
              'ESTADO pasa a «VENCIDO» solo cuando la fecha de vencimiento ya '
              'ha pasado: es el aviso de que toca reclamar.'),
        ('b', 'Si el evento llevaba barra con cobro en EFECTIVO, cuenta el '
              'dinero en la tabla opcional del final (denominaciones, menos el '
              'fondo) y, si quieres, suma tú ese efectivo neto a «Cobrado tras '
              'el evento». No se enlaza solo a propósito: en la mayoría de los '
              'eventos no hay efectivo y una fórmula fija te dejaría un 0 '
              'restando donde no debe.'),
        ('h', 'Registro de Eventos'),
        ('b', 'Una fila por evento: de un vistazo ves cuánto se facturó, qué '
              'anticipos entraron, qué queda pendiente y cuántos eventos están '
              'VENCIDOS. La fila TOTALES suma el año y cuenta los pendientes.'),
        ('h', 'Celdas editables'),
        ('b', 'Todo lo que se rellena va en VERDE: Responsable, Cuándo, ✓ '
              'Completada, Firma, los importes de la liquidación y las filas '
              'del registro de eventos.'),
    ]


def instrucciones_negocio(fname):
    kit = CTX.get('kit', '')
    areas = CTX.get('f_areas')
    caja = CTX.get('f_caja')
    titulo = 'Apertura y Cierre del Negocio' + (f' — {kit}' if kit else '')
    bloques = [
        ('h', 'Qué resuelve'),
        ('b', 'Es el checklist del LOCAL COMPLETO: lo que se abre y se cierra '
              'del negocio entero, no de cada área. Úsalo como marco del día.'),
    ]
    # m6 — en el modelo POR EVENTOS el fichero del dinero no se explica solo
    # desde «Se conecta con»: es el que sostiene la facturación del kit y su
    # nombre no dice «caja» por ningún lado. Se nombra aquí, igual que se
    # nombra el de áreas cuando lo hay. Condicionado al modelo: en los 10 kits
    # de mostrador esta línea NO se emite y sus Instrucciones no se mueven.
    if caja and es_modelo_eventos():
        bloques.append(('b', 'El DINERO no está aquí: los anticipos, la '
                             'liquidación y el cobro de cada evento van en '
                             + caja + '.'))
    if areas:
        # T-04 — mismo criterio que en «Se conecta con»: las zonas salen de los
        # nombres de hoja del fichero de áreas, no de un literal fijo.
        bloques.append(('b', 'El detalle por área'
                             + parentesis(CTX.get('areas_nombres'))
                             + f' va en {areas}: no repitas el trabajo, cada '
                             'fichero cubre un nivel distinto.'))
    bloques += [
        ('h', 'Cómo usar'),
        ('b', 'Imprime la hoja del turno (apertura o cierre) o úsala en '
              'tablet; el responsable reparte y cada persona firma su tarea.'),
        ('b', 'Cada tarea trae Responsable y Hora Límite precargados con el '
              'criterio del kit: ajústalos a tu horario real, son celdas '
              'verdes.'),
        ('h', 'Celdas editables'),
        ('b', 'Todo lo que se rellena a diario va en VERDE: Responsable, Hora '
              'Límite, ✓ Completada, Firma y Notas.'),
    ]
    return titulo, bloques


def reescribir_instrucciones(wb, fname, cambios):
    """Reconstruye la hoja entera: molde ▸, «Se conecta con», bio y versión.

    Se RECONSTRUYE (no se parchea) para que la 2.ª pasada dé exactamente el
    mismo resultado: se leen los bloques que ya había, se descartan los que
    gestiona el motor y se vuelven a emitir todos en orden fijo.
    """
    if 'Instrucciones' not in wb.sheetnames:
        return
    ws = wb['Instrucciones']
    col = 2 if any(isinstance(ws.cell(row=r, column=2).value, str)
                   for r in range(1, min(ws.max_row, 12) + 1)) else 1
    lineas = [ws.cell(row=r, column=col).value
              for r in range(1, ws.max_row + 1)
              if isinstance(ws.cell(row=r, column=col).value, str)
              and ws.cell(row=r, column=col).value.strip()]

    # CB-E4 — el bloque heredado sólo se descarta cuando el motor va a emitir
    # el suyo en su lugar. Descartarlo siempre borraría contenido bueno de los
    # kits de la familia que usan el mismo encabezado.
    sustituye_personalizar = bool(
        sub_cb() and CTX.get('personalizar', {}).get(fname))
    propio = fname in (CTX.get('f_caja'), CTX.get('f_negocio'))
    if propio:
        if fname == CTX.get('f_caja'):
            # m6 — el mismo hueco, dos manuales incompatibles: arqueo de
            # mostrador o liquidación por evento. Lo decide el modelo del kit,
            # nunca el nombre del fichero (los dos se llaman «09-…»).
            titulo, bloques = (instrucciones_cobros(fname)
                               if es_modelo_eventos()
                               else instrucciones_caja(fname))
        else:
            titulo, bloques = instrucciones_negocio(fname)
    else:
        titulo = lineas[0] if lineas else fname
        bloques = []
        actual = None
        for v in lineas[1:]:
            if v.startswith('—') or RX_BIO.search(v) or RX_VERSION.match(v) \
                    or RX_CONTACTO.match(v) or RX_COPY.match(v):
                continue
            if v.startswith('▸'):
                if actual in MIS_BLOQUES or RX_OBSOLETO.search(v) \
                        or (sustituye_personalizar and actual
                            and RX_PERSONALIZAR.match(actual)):
                    continue
                bloques.append(('b', v[1:].strip()))
            else:
                actual = v.strip().rstrip(':')
                # CB-E4 — «Cómo personalizar» pasa a ser un bloque del motor:
                # se descarta el heredado (aquí) y se emite el del TIPO de
                # entregable más abajo.
                if actual in MIS_BLOQUES or v.startswith('Marca con') \
                        or (sustituye_personalizar
                            and RX_PERSONALIZAR.match(actual)):
                    continue
                bloques.append(('h', actual))
        # Un encabezado que se ha quedado sin ni una viñeta detrás (todas
        # obsoletas) no se emite: dejaría un rótulo suelto en la hoja.
        bloques = [b for i, b in enumerate(bloques)
                   if b[0] == 'b' or (i + 1 < len(bloques)
                                      and bloques[i + 1][0] == 'b')]

    if fname == CTX.get('f_plantilla'):
        bloques = _bloque_plantilla() + bloques
    # CB-E4 — sólo en la sub-familia CB y sólo donde había un «Cómo
    # personalizar» que sustituir.
    #
    # La condición del encabezado NO basta, y lo dice un dato: el barrido de
    # los 13 kits de la familia (2026-08-29) encontró que
    # `kit-tareas/01-apertura-cierre.xlsx` —el REPRESENTANTE, publicado— trae
    # exactamente ese encabezado, así que sin `sub_cb()` la extensión le
    # habría reescrito las Instrucciones. Y el gate bloqueante de §7-bis.24
    # NO lo habría visto: corre sobre cafetería y hotel, no sobre `kit-tareas`.
    if sustituye_personalizar:
        bloques += _bloque_personalizar(
            CTX.get('papel_instrucciones', {}).get(fname, 'checklist'))
    if fname in CTX.get('con_checklist', ()):
        bloques += _bloques_contador()
        bloques += _bloque_proteccion()
    bloques += _bloque_conecta(fname)

    # --- pintado ----------------------------------------------------------
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for row in ws.iter_rows():
        for c in row:
            c.value = None
    for r in list(ws.row_dimensions):
        ws.row_dimensions[r].height = None
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = ANCHO_B

    def escribe(fila, texto, fuente):
        # TEC-R2-06 — la altura tiene que salir del CUERPO de la fuente, no de
        # un 15 pt fijo: la línea de bio son 116 caracteres a 12 pt (2 líneas de
        # 16,4 = 32,7 pt) y se publicaba con 30, comiéndose los descendentes de
        # la segunda línea al imprimir.
        cel = ws.cell(row=fila, column=2, value=texto)
        cel.font = fuente
        cel.alignment = Alignment(wrap_text=True, vertical='top')
        size = fuente.size or 11
        cap = max(20, int(ANCHO_B * 11 / size))
        alto = 15 * size / 11.0
        ws.row_dimensions[fila].height = max(
            15, math.ceil(math.ceil(len(texto) / cap) * alto * 1.05))

    escribe(2, titulo, F_TITULO)
    ws.row_dimensions[2].height = 26
    fila = 4
    for tipo, texto in bloques:
        if tipo == 'h':
            if fila > 4:
                fila += 1
            escribe(fila, texto, F_CAB)
            fila += 2
        else:
            escribe(fila, '▸ ' + texto, F_TXT)
            fila += 1
    fila += 1
    if CTX.get('pie'):
        escribe(fila, CTX['pie'], F_CAB)
        fila += 2
    escribe(fila, BIO, F_CAB)
    fila += 1
    escribe(fila, 'Contacto: info@aichef.pro', F_CAB)
    fila += 2
    escribe(fila, version_line(), F_CAB)
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.print_area = f'A1:B{fila}'
    cambios.append(f'Instrucciones reescritas ({len(bloques)} bloques) con '
                   '«Se conecta con», bio anclada y versión 2.0')


def version_line():
    pid = CTX.get('producto', 'kit-tareas')
    return f'Versión 2.0 · agosto 2026 · aichef.pro/{pid} · info@aichef.pro'


# ==========================================================================
# §2.7 — protección
# ==========================================================================
def proteger(ws, cuerpo, cambios):
    """Protección SIN contraseña: se desbloquea el CUERPO ENTERO de la tabla
    (numeración, rótulos de sección y filas libres incluidos: en el 07 el
    cliente tiene que poder renombrar «APERTURA / SERVICIO / CIERRE») y toda
    celda verde; el resto —fórmulas del contador y del arqueo, cabeceras,
    pies— queda bloqueado."""
    libres = 0
    for row in ws.iter_rows():
        for c in row:
            dentro = bool(cuerpo) and cuerpo[0] <= c.row <= cuerpo[1]
            if dentro or _es_color(c, VERDE):
                c.protection = Protection(locked=False)
                libres += 1
            else:
                c.protection = Protection(locked=True)
    ws.protection.sheet = True
    # SIN contraseña: no se toca `password` (asignar '' escribe el hash de la
    # cadena vacía y Excel pediría contraseña al desproteger).
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False          # False = PERMITIDO insertar
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    return libres


# ==========================================================================
# Metadata
# ==========================================================================
#: m5 — firma de la convención de `keywords` de la familia: «<términos del
#: kit>, AI Chef Pro». Es lo único que se exige, y por un motivo medido: los 15
#: ficheros de `kit-tareas-pasteleria` llevan «pastelería, obrador, checklist,
#: tareas, AI Chef Pro» —escrito a mano, más rico que cualquier cosa derivable
#: del identificador— y una escritura ABSOLUTA lo habría sustituido por «kit
#: tareas pasteleria, AI Chef Pro», borrando metadatos buenos en un producto
#: que ni siquiera es de esta familia (tiene su propio postproceso). Se escribe
#: sólo cuando falta o cuando no sigue la convención.
COLA_KEYWORDS = 'AI Chef Pro'


def keywords_del_kit():
    """m5 — `keywords` por defecto, DERIVADO del identificador del producto.

    Se deriva y no se inventa porque tiene que reproducir exactamente lo que ya
    hay en producción: `kit-tareas` → «kit tareas, AI Chef Pro»,
    `kit-tareas-restaurante-creativo` → «kit tareas restaurante creativo, AI
    Chef Pro» (verificado en las 19 carpetas `kit-tareas*` de `dl/`: 206 de 221
    ficheros ya lo cumplen al carácter). Una cadena nueva cambiaría los 11
    ficheros del representante, que la regresión exige idénticos.
    """
    pid = CTX.get('producto') or 'kit-tareas'
    return pid.replace('-', ' ') + ', ' + COLA_KEYWORDS


def keywords_ok(v):
    """¿Estas `keywords` siguen la convención de la familia? (m5)"""
    return bool(v) and v.strip().endswith(COLA_KEYWORDS)


def _nombre_documento(wb):
    """m5 — nombre humano del documento: el título de «Instrucciones» o, si no
    lo hay, la primera celda con texto de la primera hoja de datos."""
    if 'Instrucciones' in wb.sheetnames:
        ws = wb['Instrucciones']
        for r in (2, 1, 3):
            for c in (2, 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip():
                    return v.strip()
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        for c in (1, 2):
            v = ws.cell(row=1, column=c).value
            if isinstance(v, str) and v.strip():
                return v.strip()
    return None


def set_metadata(wb, fname, cambios):
    """m5 — metadata coherente en CUALQUIER fichero del producto.

    Hasta el motor 2.4 esto vivía dentro de `cerrar()`, que sale antes de
    llamarlo cuando el fichero está FUERA del molde ▸. Resultado: los ficheros
    del molde P4 y los dos BONUS de los cinco kits con alcance «sólo 08/09» se
    guardaban (reciben desplegable, contador, bio y versión 2.0 desde
    `normalizar_p4`) pero conservaban `subject` «… · v1.1». El cliente abría un
    producto v2.0 y las propiedades de la mayoría de sus ficheros decían otra
    versión. Ahora lo llama `main.procesar` para TODOS los ficheros, después de
    `cerrar` — que es quien reescribe «Instrucciones», de donde sale el título.

    El `title` se recompone SÓLO si no está ya en la forma canónica
    «<nombre> · <sufijo>» (COM-27: había títulos genéricos «Kit de Tareas — …»).
    Reescribirlo siempre no aportaría nada y pondría en riesgo los 11 títulos
    del representante, que la regresión exige byte a byte.
    """
    p = wb.properties
    antes = (p.title, p.subject, p.creator, p.keywords)
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'
    sufijo = CTX.get('sufijo') or 'Kit de Tareas Recurrentes Pro'
    p.subject = f'{sufijo} · v2.0'
    if not keywords_ok(p.keywords):
        p.keywords = keywords_del_kit()
    titulo = p.title or ''
    if not titulo.endswith(' · ' + sufijo):
        nombre = _nombre_documento(wb)
        if nombre:
            p.title = f'{nombre} · {sufijo}'
    if (p.title, p.subject, p.creator, p.keywords) != antes:
        cambios.append(
            'metadata: title «{}» · subject «{}» · keywords «{}»'.format(
                p.title, p.subject, p.keywords))



# ==========================================================================
# CB-E1 · ORTOGRAFÍA — tildes y ñ (sub-familia ChefBusiness)
# ==========================================================================
#: Los siete kits portados de ChefBusiness. Es una DECLARACIÓN explícita, no
#: una heurística, y el motivo está medido: después de CB-E1 no queda ninguna
#: firma estructural que los separe de la familia ▸ (misma cabecera de siete
#: columnas, mismos rótulos; lo único que hoy los distingue —«Hora Limite» sin
#: tilde— es justamente lo que este paso corrige, así que una detección por esa
#: vía dejaría de funcionar en la 2.ª pasada). Y el precio de equivocarse es
#: alto: `motor.py` sostiene 11 kits en producción y CB-E9 (anchos, wrap, verde
#: del CF) tiene la MISMA firma de defecto en cafetería (446 celdas de tarea sin
#: `wrap_text`), en pizzería (319), en hamburguesería (292) y en dark-kitchen
#: (275) — kits ya publicados que §7-bis.24 congela. Sin esta lista, «arreglar
#: la legibilidad» habría reescrito cuatro productos que nadie ha pedido tocar.
SUBFAMILIA_CB = frozenset((
    'kit-tareas-sushi-bar', 'kit-tareas-asador', 'kit-tareas-marisqueria',
    'kit-tareas-panaderia', 'kit-tareas-food-truck', 'kit-tareas-tapas-bar',
    'kit-tareas-chef-privado'))


def sub_cb():
    """¿El producto en curso es de la sub-familia ChefBusiness? (CB-E1/E9)"""
    return CTX.get('producto') in SUBFAMILIA_CB


#: Diccionario EXPLÍCITO de lemas mal escritos → forma correcta, en minúsculas.
#: No es heurística: cada entrada se ha derivado del propio corpus (la misma
#: palabra escrita con y sin tilde dentro de los 20 kits de `dl/`, medido el
#: 2026-08-29) y se ha filtrado a mano. Fuera quedan, a propósito:
#:  · los homógrafos donde la forma SIN tilde también es palabra («el/él»,
#:    «esta/está», «mas/más», «solo/sólo», «como/cómo», «si/sí», «tu/tú»,
#:    «te/té», «uso/usó», «paso/pasó», «cambio/cambió», «quien/quién»,
#:    «donde/dónde», «critico/crítico», «perdida/pérdida», «publico/público»);
#:  · «carnes» (que NO es «carnés»), «min»/«max» (abreviaturas legítimas),
#:    «periodo» (variante válida) y «celiaco» (variante válida en el DLE);
#:  · **«campana»/«campanas»**, que en hostelería es la CAMPANA extractora y
#:    está bien escrita: el gemelo «campaña» del corpus es de otro contexto.
#:    Traducirla habría escrito «campaña extractora» en `kit-tareas-hotel`
#:    (`13-mantenimiento.xlsx:'Mensual'!B6`), un kit LIVE.
#: Las ambigüedades no se resuelven aquí: el gate las EMITE para que el módulo
#: de contenido de cada kit las decida celda a celda (§1.2 CB-E1).
LEX_TILDES = {
        'accion': 'acción', 'acompanamiento': 'acompañamiento',
        'acompanar': 'acompañar', 'ademas': 'además',
        'administracion': 'administración', 'albaran': 'albarán',
        'alergeno': 'alérgeno', 'alergenos': 'alérgenos',
        'alimentacion': 'alimentación', 'almacen': 'almacén', 'anade': 'añade',
        'anadido': 'añadido', 'anadir': 'añadir', 'analisis': 'análisis',
        'angulo': 'ángulo', 'ano': 'año', 'anos': 'años',
        'antelacion': 'antelación', 'aqui': 'aquí', 'area': 'área',
        'articulo': 'artículo', 'atencion': 'atención', 'atun': 'atún',
        'auditoria': 'auditoría', 'autonoma': 'autónoma', 'azucar': 'azúcar',
        'bano': 'baño', 'banos': 'baños', 'bascula': 'báscula',
        'basculas': 'básculas', 'basica': 'básica', 'bolleria': 'bollería',
        'boqueron': 'boquerón', 'cafeteria': 'cafetería', 'cajon': 'cajón',
        'calabacin': 'calabacín', 'calefaccion': 'calefacción',
        'calibracion': 'calibración', 'camara': 'cámara', 'camaras': 'cámaras',
        'canapes': 'canapés', 'castanas': 'castañas', 'categoria': 'categoría',
        'categorias': 'categorías', 'champan': 'champán',
        'chocolateria': 'chocolatería', 'circulacion': 'circulación',
        'coccion': 'cocción', 'cocteleria': 'coctelería',
        'companeros': 'compañeros', 'congelacion': 'congelación',
        'consultoria': 'consultoría', 'corazon': 'corazón',
        'cristaleria': 'cristalería', 'crustaceos': 'crustáceos',
        'cuberteria': 'cubertería', 'datafono': 'datáfono',
        'decision': 'decisión', 'decoracion': 'decoración',
        'degustacion': 'degustación', 'deposito': 'depósito',
        'desague': 'desagüe', 'desagues': 'desagües',
        'descongelacion': 'descongelación', 'desempeno': 'desempeño',
        'desinfeccion': 'desinfección', 'despues': 'después', 'dia': 'día',
        'dias': 'días', 'direccion': 'dirección', 'disenado': 'diseñado',
        'disenar': 'diseñar', 'diseno': 'diseño',
        'documentacion': 'documentación', 'donacion': 'donación',
        'duracion': 'duración', 'ejecucion': 'ejecución',
        'elaboracion': 'elaboración', 'electrica': 'eléctrica',
        'espana': 'españa', 'espanol': 'español', 'espanola': 'española',
        'esparragos': 'espárragos', 'espatula': 'espátula',
        'espatulas': 'espátulas', 'especifica': 'específica',
        'especificas': 'específicas', 'especifico': 'específico',
        'especificos': 'específicos', 'estacion': 'estación', 'estan': 'están',
        'estandar': 'estándar', 'estandares': 'estándares',
        'evacuacion': 'evacuación', 'evaluacion': 'evaluación',
        'exposicion': 'exposición', 'extraccion': 'extracción',
        'facturacion': 'facturación', 'fermentacion': 'fermentación',
        'fidelizacion': 'fidelización', 'formacion': 'formación',
        'frias': 'frías', 'frigorificas': 'frigoríficas', 'frio': 'frío',
        'frios': 'fríos', 'futbol': 'fútbol', 'gastronomica': 'gastronómica',
        'gastronomicas': 'gastronómicas', 'gastronomicos': 'gastronómicos',
        'gestion': 'gestión', 'hamburgueseria': 'hamburguesería',
        'heladeria': 'heladería', 'hermetico': 'hermético',
        'higienico': 'higiénico', 'higienicos': 'higiénicos', 'hollin': 'hollín',
        'hosteleria': 'hostelería', 'humedo': 'húmedo',
        'iluminacion': 'iluminación', 'informacion': 'información',
        'inspeccion': 'inspección', 'jabon': 'jabón', 'jamon': 'jamón',
        'lacteos': 'lácteos', 'laminas': 'láminas', 'lena': 'leña',
        'limite': 'límite', 'limites': 'límites', 'limon': 'limón',
        'linea': 'línea', 'lineas': 'líneas', 'liquidos': 'líquidos',
        'logistica': 'logística', 'maduracion': 'maduración', 'manana': 'mañana',
        'manipulacion': 'manipulación', 'marabu': 'marabú', 'maxima': 'máxima',
        'maximo': 'máximo', 'menu': 'menú', 'mercancia': 'mercancía',
        'miercoles': 'miércoles', 'minima': 'mínima', 'minimo': 'mínimo',
        'movil': 'móvil', 'musica': 'música', 'navidenos': 'navideños',
        'neumaticos': 'neumáticos', 'nino': 'niño', 'ninos': 'niños',
        'numero': 'número', 'numeros': 'números', 'optimo': 'óptimo',
        'otono': 'otoño', 'panaderia': 'panadería', 'pano': 'paño',
        'panos': 'paños', 'participacion': 'participación',
        'pasteleria': 'pastelería', 'pequeno': 'pequeño', 'pequenos': 'pequeños',
        'periodica': 'periódica', 'periodico': 'periódico',
        'pizzeria': 'pizzería', 'planificacion': 'planificación',
        'portatil': 'portátil', 'posicion': 'posición', 'precision': 'precisión',
        'preparacion': 'preparación', 'presentacion': 'presentación',
        'presion': 'presión', 'prevision': 'previsión',
        'produccion': 'producción', 'promocion': 'promoción',
        'proteina': 'proteína', 'proteinas': 'proteínas', 'proxima': 'próxima',
        'proximo': 'próximo', 'proximos': 'próximos', 'quimicos': 'químicos',
        'racion': 'ración', 'rapida': 'rápida', 'rapido': 'rápido',
        'recepcion': 'recepción', 'refrigeracion': 'refrigeración',
        'renovacion': 'renovación', 'repeticion': 'repetición',
        'reposicion': 'reposición', 'resenas': 'reseñas', 'reunion': 'reunión',
        'revision': 'revisión', 'rigida': 'rígida', 'romantica': 'romántica',
        'roscon': 'roscón', 'rotacion': 'rotación', 'sabado': 'sábado',
        'salmon': 'salmón', 'sandia': 'sandía', 'seccion': 'sección',
        'segun': 'según', 'seleccion': 'selección',
        'senal': 'señal', 'senales': 'señales', 'senaletica': 'señalética',
        'senalizacion': 'señalización', 'sesamo': 'sésamo', 'sesion': 'sesión',
        'sifon': 'sifón', 'simbolos': 'símbolos', 'supervision': 'supervisión',
        'tambien': 'también', 'tecnica': 'técnica', 'tecnicas': 'técnicas',
        'tecnico': 'técnico', 'tematica': 'temática', 'tematico': 'temático',
        'termometro': 'termómetro', 'termometros': 'termómetros',
        'transicion': 'transición', 'turron': 'turrón', 'ubicacion': 'ubicación',
        'ultima': 'última', 'ultimas': 'últimas', 'ultimo': 'último',
        'ultimos': 'últimos', 'util': 'útil', 'vacias': 'vacías',
        'vacio': 'vacío', 'vacios': 'vacíos', 'valentin': 'valentín',
        'vehiculo': 'vehículo', 'ventilacion': 'ventilación',
    # --- Segunda vuelta (2026-08-29). El método de arriba —«la misma palabra
    # escrita con y sin tilde dentro del corpus»— no ve las que el corpus NUNCA
    # escribe bien: no tienen gemelo del que aprender. Estas 23 salieron de un
    # barrido por FAMILIA DE SUFIJO (-ción, -sión, -metro, -il/-iles) sobre los
    # siete kits, y una de ellas, «Operaciones Moviles», es una pestaña que la
    # SPEC ya había listado a mano en §0.3 y que el diccionario derivado del
    # corpus se dejaba fuera.
    # ⚠️ Los PLURALES en -ciones y -siones NO llevan tilde («instrucciones»,
    # «guarniciones», «observaciones», «operaciones», «condiciones»,
    # «restricciones»…): el mismo barrido los señaló y son correctos tal cual.
    # Meterlos habría escrito «instruccíones» en las 38 pestañas del kit.
    'acumulacion': 'acumulación', 'anticipacion': 'anticipación',
    'clasificacion': 'clasificación', 'division': 'división',
    'estimacion': 'estimación', 'extension': 'extensión',
    'flotacion': 'flotación', 'hidratacion': 'hidratación',
    'importacion': 'importación', 'inversion': 'inversión',
    'legislacion': 'legislación', 'lubricacion': 'lubricación',
    'moviles': 'móviles', 'organizacion': 'organización',
    'oxigenacion': 'oxigenación', 'pirometro': 'pirómetro',
    'pirometros': 'pirómetros', 'porcion': 'porción',
    'precoccion': 'precocción', 'recomendacion': 'recomendación',
    'recuperacion': 'recuperación', 'regulacion': 'regulación',
    'solucion': 'solución', 'sustitucion': 'sustitución',
}

#: Frases fijas donde «como» sí lleva tilde: son encabezados del molde, no
#: lenguaje libre, así que se pueden tratar por literal sin ambigüedad.
LEX_FRASES = {
    'como usar': 'cómo usar',
    'como personalizar': 'cómo personalizar',
    'como cuenta': 'cómo cuenta',
    'como se': 'cómo se',
}

#: Un token que contenga esto es una URL, un correo, un dominio o un NOMBRE DE
#: FICHERO: no se toca.
#: Va por TOKEN y no por celda, y la diferencia tiene nombre y apellidos: el pie
#: de los seis kits de CB es «— Kit de Tareas: Sushi Bar · ChefBusiness
#: Consultoria Gastronomica · chefbusiness.co», o sea la razón social mal
#: escrita 11-13 veces por kit EN LA MISMA CELDA que el dominio. Excluyendo la
#: celda entera (que fue el primer diseño) §7-bis.17 se quedaba sin aplicar.
#: La regla es «un punto entre dos alfanuméricos», y no una lista de dominios,
#: por un falso positivo MEDIDO: con la lista, «Se conecta con» de catering
#: —un kit LIVE— pasaba a citar «09-cobros-facturación-eventos.xlsx», un
#: fichero que no existe. El nombre del fichero es una referencia, no prosa.
#: La barra NO entra en la regla: «Lote/Albaran» y «Proveedor/Lonja» son
#: rótulos de tabla y sí hay que corregirlos.
RX_TOKEN_URL = re.compile(r'(://|@|[A-Za-z0-9]\.[A-Za-z0-9])')
#: Siglas y códigos: se dejan intactos aunque casen con un lema. En mayúsculas
#: y de 2 a 5 letras («FAO», «RD», «CE», «UE», «APPCC», «IVA», «TPV»). «ANO» y
#: «OTONO» no caen aquí: tienen 3 y 5 letras pero están en `LEX_TILDES` y la
#: comprobación de sigla sólo se aplica a lo que NO está en el diccionario.
RX_SIGLA = re.compile(r'^[A-ZÁÉÍÓÚÜÑ]{2,5}$')
RX_PALABRA = re.compile(r'[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+')


def _caso(origen, destino):
    """Aplica a `destino` el caso de `origen` (minúscula / Título / VERSALITA)."""
    if origen.isupper() and len(origen) > 1:
        return destino.upper()
    if origen[:1].isupper():
        return destino[:1].upper() + destino[1:]
    return destino


def _orto_token(tok):
    """Corrige UN token (trozo sin espacios). Devuelve (nuevo, nº de cambios)."""
    if RX_TOKEN_URL.search(tok):
        return tok, 0
    n = 0

    def _rep(m):
        nonlocal n
        p = m.group(0)
        destino = LEX_TILDES.get(p.lower())
        if not destino or destino == p.lower():
            return p
        if RX_SIGLA.match(p) and p.lower() not in LEX_TILDES:
            return p
        nuevo = _caso(p, destino)
        if nuevo != p:
            n += 1
        return nuevo

    return RX_PALABRA.sub(_rep, tok), n


def orto_texto(v):
    """Corrige un texto entero. Devuelve (nuevo, nº de palabras corregidas)."""
    if not isinstance(v, str) or not v.strip():
        return v, 0
    total = 0
    # 1) frases fijas (el «Cómo» de los encabezados del molde)
    for mal, bien in LEX_FRASES.items():
        rx = re.compile(r'(?i)\b' + re.escape(mal) + r'\b')

        def _f(m, bien=bien):
            nonlocal total
            partes = m.group(0).split(' ')
            arreglo = [_caso(a, b) for a, b in zip(partes, bien.split(' '))]
            nuevo = ' '.join(arreglo)
            if nuevo != m.group(0):
                total += 1
            return nuevo
        v = rx.sub(_f, v)
    # 2) palabra a palabra, token a token (los tokens de URL quedan fuera)
    fuera = []
    for tok in re.split(r'(\s+)', v):
        if tok.strip():
            tok, n = _orto_token(tok)
            total += n
        fuera.append(tok)
    return ''.join(fuera), total


#: Literales que son CONTRATO entre el motor y el fichero: los reconoce
#: `geometria`, `_contador`, `cadencia` o la DV, y reescribirlos rompería la
#: detección. Se comparan por valor exacto de celda.
def _contrato():
    return {DV_LISTA, CAB_MARCA, CAB_P4, ETIQ_CONTADOR, CAB_EFECTIVO,
            ETIQ_FONDO, ETIQ_FONDO_RESUMEN, ETIQ_EFECTIVO, ETIQ_VENTAS_EF,
            ETIQ_Z, MARCA_OK, MARCA_NO} | set(CAB_REGISTRO) | set(
                CAB_RECUENTO) | set(CAB_CALENDARIO) | set(CAB_CAJA) | set(
                    CAB_EVENTOS)


def ortografia(wb, fname, cambios):
    """CB-E1 — barrido de tildes y ñ. Paso PROPIO, DESPUÉS de todo lo demás.

    El orden no es una preferencia: está medido. `reescribir_instrucciones`
    RECONSTRUYE la hoja leyendo las líneas que ya había, así que reinyecta el
    texto sin tilde del fichero original — tras el dry-run del 2026-08-29,
    `01-apertura-cierre-sushi.xlsx:Instrucciones!B4` seguía diciendo «Como usar
    estas plantillas». Corriendo antes, el motor lo desharía (§7-bis.11).

    Y por eso mismo hay una condición que este paso IMPONE al resto del motor:
    ningún detector puede depender de una tilde. Si `_col_tiempo` sólo conociera
    «Hora Límite», la 1.ª pasada no vería la columna de tiempo de sushi-bar
    («Hora Limite») y la 2.ª —ya corregida— sí: cabecera reescrita, subtítulo
    reescrito, columna pintada de verde y el gate de idempotencia en rojo sin
    nada roto de verdad. Por eso `_col_tiempo` y `EDITABLES` comparan sin
    tildes desde esta misma tanda.

    No toca: fórmulas (`data_type == 'f'`), tokens con URL / correo / dominio,
    siglas de 2-5 letras que no estén en el diccionario, ni los literales de
    `_contrato()`. Sí toca los NOMBRES DE PESTAÑA y el `title`/`subject` de la
    metadata (no las `keywords`: son el índice del producto y `keywords_ok` las
    da por buenas tal cual).
    """
    if not sub_cb():
        # Fuera de la sub-familia el paso NO corre. No es prudencia genérica:
        # el barrido encuentra 2 faltas REALES en `kit-tareas-hamburgueseria`
        # («sesamo» en `02-partidas-cocina.xlsx:'Línea Montaje'!B9` y en
        # `05-tareas-semanales-mensuales.xlsx:'Inventario Diario'!A13`), que es
        # un kit LIVE y no está en esta tanda. Corregirlas de paso lo habría
        # modificado sin diff firmado, sin regresión propia y sin APPLY — justo
        # lo que §7-bis.24 viene a impedir. Quedan declaradas en el informe.
        return {'palabras': 0, 'celdas': 0, 'hojas': 0, 'fuera_de_alcance': 1}
    contrato = _contrato()
    celdas = hojas = 0
    palabras = 0
    for ws in list(wb.worksheets):
        nuevo, n = orto_texto(ws.title)
        if n and nuevo not in wb.sheetnames:
            viejo = ws.title
            ws.title = nuevo
            hojas += 1
            palabras += n
            # El REGISTRO de fórmulas guarda el TÍTULO de la hoja, y `main.py`
            # lo usa después para comprobar una a una que quedaron con valor
            # cacheado. Renombrar la pestaña sin actualizarlo dejaba el gate
            # buscando «Temporadas Pescado Espana» en un libro donde ya se
            # llamaba «…España»: dos fórmulas del 08 de sushi-bar salían como
            # «hoja ausente» con la caché perfectamente inyectada.
            for i, (t, coord, f) in enumerate(REGISTRO):
                if t == viejo:
                    REGISTRO[i] = (nuevo, coord, f)
            cambios.append(f'CB-E1: pestaña «{viejo}» → «{nuevo}»')
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or c.data_type == 'f' \
                        or v.startswith('='):
                    continue
                if v in contrato:
                    continue
                nuevo, n = orto_texto(v)
                if n and nuevo != v:
                    c.value = nuevo
                    celdas += 1
                    palabras += n
    p = wb.properties
    for campo in ('title', 'subject'):
        v = getattr(p, campo, None)
        if isinstance(v, str):
            nuevo, n = orto_texto(v)
            if n and nuevo != v:
                setattr(p, campo, nuevo)
                palabras += n
    if palabras:
        cambios.append(f'CB-E1 ortografía: {palabras} palabras con tilde o ñ '
                       f'restituidas en {celdas} celdas'
                       + (f' y {hojas} pestañas' if hojas else ''))
    return {'palabras': palabras, 'celdas': celdas, 'hojas': hojas}


def palabras_sin_tilde(wb):
    """Gate `ortografia`: lemas de `LEX_TILDES` que sobreviven. Debe ser 0."""
    fuera = []
    contrato = _contrato()
    for ws in wb.worksheets:
        _, n = orto_texto(ws.title)
        if n:
            fuera.append((ws.title, '·hoja·', ws.title))
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or c.data_type == 'f' \
                        or v.startswith('=') or v in contrato:
                    continue
                _, n = orto_texto(v)
                if n:
                    fuera.append((ws.title, c.coordinate, v))
    return fuera


#: CB-E1 (segunda mitad) — las AMBIGUAS. No se corrigen por diccionario: el
#: gate las lista con su celda para que el módulo de contenido del kit las
#: decida una a una. Cada pareja es (forma sin tilde, forma con tilde).
AMBIGUAS = {
    'el': 'él', 'esta': 'está', 'estas': 'estás', 'este': 'esté',
    'mas': 'más', 'si': 'sí', 'solo': 'sólo', 'tu': 'tú', 'te': 'té',
    'como': 'cómo', 'cuando': 'cuándo', 'donde': 'dónde', 'quien': 'quién',
    'que': 'qué', 'uso': 'usó', 'paso': 'pasó', 'cambio': 'cambió',
    'critico': 'crítico', 'criticas': 'críticas', 'perdida': 'pérdida',
    'publico': 'público', 'publica': 'pública', 'min': 'mín', 'max': 'máx',
}
RX_AMBIGUA = re.compile(r'(?i)\b(' + '|'.join(sorted(AMBIGUAS)) + r')\b')


def ambiguas_del_libro(wb):
    """Ocurrencias de palabras AMBIGUAS, para que el contenido las decida."""
    fuera = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or c.data_type == 'f':
                    continue
                for m in RX_AMBIGUA.finditer(v):
                    fuera.append({'hoja': ws.title, 'celda': c.coordinate,
                                  'palabra': m.group(0), 'texto': v[:120]})
    return fuera


# ==========================================================================
# API principal
# ==========================================================================
def aplicar(wb, fname, cambios):
    """§1 y §2 sobre un libro ya cargado. Devuelve las hojas tocadas."""
    recon = hojas_reconocidas(wb)
    if not recon:
        # R3-e — el fichero no es del molde ▸, pero puede ser del molde P4
        # (los 01-07 de catering, chocolatería, heladería, hotel y
        # restaurante-creativo). Se le pasa la normalización mínima para que el
        # producto no entregue dos desplegables y dos contadores distintos.
        normalizar_p4(wb, cambios)
        return {}

    # DV y CF se vacían y se reconstruyen enteros (idempotencia)
    for titulo in recon:
        ws = wb[titulo]
        ws.data_validations.dataValidation = []
        ws.conditional_formatting = ConditionalFormattingList()

    es_caja = fname == CTX.get('f_caja')
    es_negocio = fname == CTX.get('f_negocio')
    # m6 — el fichero del dinero puede ser de MOSTRADOR (arqueo) o de EVENTOS
    # (liquidación). Todo lo que sigue tocando cajón, TPV y fondo es del
    # primero: aplicárselo al segundo escribiría «Responsable de caja» y horas
    # de reloj encima de los responsables y de los D-15/D-7 del diseño.
    es_mostrador = es_caja and not es_modelo_eventos()
    es_cobros = es_caja and es_modelo_eventos()

    # 0) T-02 — el TPV se enciende UNA vez en el kit, y es en el fichero de
    #    negocio. En el de caja la tarea pasa a «Comprobar que el TPV está
    #    encendido y abrir turno de caja». Va antes que nada porque cambia el
    #    TEXTO que luego miden `autoaltos` y `anotar_duplicados`.
    if es_mostrador:
        for titulo, tipo in recon.items():
            if tipo == 'checklist':
                tpv_de_caja(wb[titulo], cambios)

    # 1) cambios de COLUMNA (antes de medir geometrías)
    if es_mostrador:
        for titulo in ('Apertura de Caja', 'Cierre de Caja'):
            if titulo in wb.sheetnames:
                caja_columnas(wb[titulo], cambios)
        if 'Registro Mensual' in wb.sheetnames:
            registro_mensual(wb['Registro Mensual'], cambios)

    # 2) borrado de filas duplicadas (§2.5). NUNCA en los ficheros de negocio y
    #    de caja: son el DESTINO de la remisión y colapsar allí un bloque
    #    homónimo lo dejaría apuntando a sí mismo.
    if not (es_caja or es_negocio):
        for titulo in list(recon):
            if recon[titulo] == 'checklist':
                colapsar_duplicados(wb[titulo], cambios)
        # R3-f — y las bandas que duplican el marco sin llamarse como en el
        # representante se ANOTAN (no se borran) en el fichero de ÁREAS, que es
        # el «01» de la jerarquía.
        if fname == CTX.get('f_areas'):
            for titulo in list(recon):
                if recon[titulo] == 'checklist':
                    anotar_duplicados(wb[titulo], cambios)

    # 2bis) CB-E2 — crear la fila de totales ANTES de normalizar: a partir de
    #    ahí `geometria` ya la encuentra y la maquinaria de siempre (5 filas
    #    libres dentro del rango, COUNTIFS y denominador honesto) hace el resto.
    for titulo, tipo in recon.items():
        if tipo == 'checklist':
            crear_contador(wb[titulo], fname, cambios)

    # 3) normalización de los checklists (§2.1/§2.2/§1.5)
    cuerpos = {}
    for titulo, tipo in recon.items():
        if tipo != 'checklist':
            continue
        g = normalizar_checklist(wb[titulo], cambios)
        if g:
            cuerpos[titulo] = (g['hr'] + 1, g['fin'])

    # 4) precargas y contenidos de la familia
    for titulo, tipo in recon.items():
        ws = wb[titulo]
        if tipo == 'checklist':
            if es_negocio:
                precargar_negocio(ws, cambios)
            elif es_mostrador:
                precargar_caja(ws, cambios)
            diferenciar_07(ws, cambios)
        elif tipo == 'calendario':
            cuerpo = calendario(ws, cambios)
            # Sin esto `proteger` sólo desbloquea las celdas verdes y el
            # calendario de bar —que no tiene ninguna— se publicaba con la
            # hoja entera bloqueada, en el fichero cuya propia portada dice
            # «Añade las fechas locales de tu zona».
            if cuerpo:
                cuerpos[titulo] = cuerpo
        elif tipo == 'briefing':
            briefing(ws, cambios)
        elif tipo == 'registro_appcc':
            # CB-E3 — formatos de fecha y temperatura, desplegable de
            # verificación, CF de fuera-de-rango y alturas de cabecera.
            cuerpo = registro_appcc(ws, cambios)
            if cuerpo:
                cuerpos[titulo] = cuerpo
        textos_de_tarea(ws, cambios, 2 if tipo == 'checklist' else None,
                        facturado=not es_cobros)
    if 'Instrucciones' in wb.sheetnames:
        textos_de_tarea(wb['Instrucciones'], cambios, facturado=not es_cobros)

    # 4bis) la cabecera de la columna de tiempo se decide con los valores YA
    #   precargados. Si se decidiese antes (como hasta ahora), la 1.ª pasada
    #   leería «Hora Límite» del original y la 2.ª el «Cierre» que escribió la
    #   precarga: dos resultados distintos y el gate de idempotencia en rojo.
    for titulo, tipo in recon.items():
        if tipo == 'checklist':
            ajustar_cabecera_tiempo(wb[titulo], cambios)

    # 4ter) alturas — R3-d: SIEMPRE después del último cambio de texto, o la
    #   2.ª pasada mide un texto distinto del que midió la 1.ª. CB-E9 (a) va
    #   JUSTO ANTES por el mismo motivo al revés: `autoalto` decide mirando
    #   `wrap_text`, así que si el ajuste de texto llegara después, la 1.ª
    #   pasada mediría la hoja sin envolver y la 2.ª con ella.
    for titulo, tipo in recon.items():
        if tipo == 'checklist':
            legibilidad_tareas(wb[titulo], cambios)
            autoaltos(wb[titulo], cambios)

    # 4quater) las hojas del libro que NO son del molde ▸ (un kit puede
    #   mezclar: el 05 de cafetería lleva inventarios) pasan por la
    #   normalización mínima del molde P4.
    normalizar_p4(wb, cambios, saltar=set(recon), bio=False)

    # 5) el arqueo (§1.1) — después de las inserciones de fila del checklist
    if es_mostrador and 'Cierre de Caja' in wb.sheetnames:
        moneda_002(wb['Cierre de Caja'], cambios)
        recuento(wb['Cierre de Caja'], cambios)
        fila_fondo = fondo_de_caja(wb, cambios)
        resumen_cierre(wb, fila_fondo, cambios)

    # 6) m8 — el formato condicional de las dos hojas del modelo POR EVENTOS.
    #    Va AL FINAL, después de cualquier inserción de fila, y es obligatorio
    #    porque el vaciado de `conditional_formatting` de arriba se lleva por
    #    delante el ámbar del PENDIENTE y el rojo del VENCIDO que trae el
    #    fichero recién construido. Sin esto el 09 de catering se publicaba
    #    sin un solo aviso de color y el dry-run seguía en verde.
    reglas_ev = 0
    for titulo, tipo in recon.items():
        if tipo in ('liquidacion', 'registro_eventos'):
            reglas_ev += cf_eventos(wb[titulo], tipo)
    if reglas_ev:
        cambios.append(f'{reglas_ev} reglas de formato condicional del modelo '
                       'por eventos (ámbar si queda PENDIENTE, rojo si el '
                       'saldo está VENCIDO)')

    return {'hojas': recon, 'cuerpos': cuerpos}


def cerrar(wb, fname, estado, cambios):
    """Instrucciones, impresión, protección y metadata."""
    if not estado:
        return
    recon, cuerpos = estado['hojas'], estado['cuerpos']
    reescribir_instrucciones(wb, fname, cambios)
    protegidas = 0
    for titulo, tipo in recon.items():
        ws = wb[titulo]
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    c.value = None                 # censo: empty_str = 0
                elif (CTX.get('pie') and isinstance(c.value, str)
                        and RX_PIE.match(c.value) and c.value != CTX['pie']):
                    # DOM-21 (f): 08 y 09 firmaban «© AI Chef Pro — aichef.pro»
                    # y los 01-07 «— Kit de Tareas Recurrentes · <kit> · …».
                    # DOM-R2-21: los dos BONUS firmaban sin el nombre del kit y
                    # uno con un guion suelto al final. Se reescribe cualquier
                    # cosa que se PAREZCA a un pie, no un literal concreto.
                    c.value = CTX['pie']
        hr = None
        g = geometria(ws)
        if g:
            hr = g['hr']
        elif fila_registro_mensual(ws):
            hr = fila_registro_mensual(ws)
        elif fila_registro_eventos(ws):
            hr = fila_registro_eventos(ws)          # m6
        elif fila_calendario(ws):
            hr = fila_calendario(ws)
        elif fila_registro_appcc(ws)[0]:
            hr = fila_registro_appcc(ws)[0]            # CB-E3
        # TEC-19: apaisado en las hojas anchas (las de 8 columnas de 08
        # escalaban al ~60 %). Nunca al revés: «Calendario Anual» y «Registro
        # Mensual» ya venían en apaisado y forzarles vertical las estropearía.
        # TEC-R2-12: el umbral es el ANCHO en unidades, no el número de
        # columnas — las dos hojas de caja miden 137-142 en 7 columnas y se
        # quedaban en vertical escalando al ~60 %, en la hoja del dinero.
        print_setup(ws, hr, landscape=(ws.max_column >= 8 or ancho_util(ws)
                                       >= 130 or ws.page_setup.orientation
                                       == 'landscape'))
        area_impresion(ws)
        # TEC-R2-14: se protegen TODAS las hojas de datos, también las de los
        # dos BONUS —el calendario ya tiene su bloque editable propio—, para
        # que §6 («hojas protegidas con las entradas desbloqueadas») sea cierto
        # de punta a punta. Las verdes quedan desbloqueadas, así que el cliente
        # sigue pudiendo escribir en su plantilla; sólo cambia que no se pisen
        # las cabeceras por accidente. Las «Instrucciones» no se protegen: son
        # texto y el cliente puede querer anotar ahí.
        # m6 — «Liquidación del Evento» y «Registro de Eventos» se protegen
        # como el resto de hojas de datos: sus celdas VERDES (los importes que
        # escribe el usuario y las 25 filas del registro) quedan desbloqueadas
        # por `proteger`, y las fórmulas del IVA, el total, el saldo, el
        # pendiente y el ESTADO, bloqueadas.
        if tipo in ('checklist', 'registro', 'registro_eventos', 'liquidacion',
                    'briefing', 'calendario', 'registro_appcc'):
            protegidas += 1
            proteger(ws, cuerpos.get(titulo), cambios)
    if protegidas:
        cambios.append(f'{protegidas} hojas protegidas sin contraseña (de '
                       f'{len(recon)} con print_area y A4)')
    # m5 — `set_metadata` YA NO se llama aquí: `cerrar` sólo corre para los
    # ficheros del molde ▸ y por eso los P4 y los BONUS se quedaban con el
    # subject de la v1.1. Lo llama `main.procesar` para TODOS los ficheros,
    # justo después de esta función.
