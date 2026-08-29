#!/usr/bin/env python3
"""
regresion.py — Comparador celda a celda de dos carpetas de entregables .xlsx.

    python3 regresion.py --a <carpeta A> --b <carpeta B> [--json informe.json]
                         [--max N] [--solo 01,08]

Nace del gate BLOQUEANTE de §7-bis.24 / `kit-tareas-cb-v2-SPEC.md` §9 T0.3: el
motor de `kit-tareas-v2_0/` SOSTIENE 11 KITS EN PRODUCCIÓN, así que toda
extensión tiene que probarse en `--dry-run` sobre dos kits ya publicados
(`kit-tareas-cafeteria` y `kit-tareas-hotel`) y dar **0 diferencias** contra
`astro-site/public/dl/<pid>/`. Sin un comparador propio, «0 diferencias» se
apoyaba en el gate de idempotencia de `main.py`, que compara la 1.ª pasada con
la 2.ª —o sea, el motor NUEVO contra sí mismo— y por construcción no puede ver
una regresión respecto de lo que hay publicado.

Qué compara, por celda (lo que pide el gate, ni más ni menos):

  · **valor** y **fórmula**  — `repr(cell.value)` + `cell.data_type`. Se carga
    con `data_only=False` a propósito: así una fórmula se compara como fórmula
    («=COUNTIF(...)») y no como el número que `inject_cache.py` le dejó en la
    caché, que cambia con los datos de ejemplo y no con el motor.
  · **number_format**
  · **locked** — de TODAS las celdas del rango usado, también las vacías: la
    protección se decide celda a celda y en BONUS-02 del representante la fila
    27 está vacía y bloqueada (m3 de `main.py:digest`).
  · **DV** (validación de datos) — tipo, fórmula, rango y **los tres textos que
    el cliente ve**: `errorTitle`, `error`, `prompt`. El mismo punto ciego que
    m3 corrigió en la huella de `main.py`.
  · **altura de fila** — `row_dimensions[r].height` de toda fila que exista en
    cualquiera de los dos lados (`autoalto` las escribe explícitas).

Y, fuera de la celda, lo que haría inútil cualquier comparación de celdas: qué
FICHEROS y qué HOJAS hay a cada lado, y en qué orden.

Lo que NO compara (y por qué): rellenos, fuentes, anchos de columna, formato
condicional, merges y metadata. No están en la lista del gate; `main.py:digest`
ya cubre relleno, merges y CF en el gate de idempotencia, y meterlos aquí
convertiría el informe en ruido antes de que nadie lo lea. Si algún día hacen
falta, entran como una dimensión más de `_celdas()`.

Salida: `exit 0` si no hay diferencias, `exit 1` si hay alguna, `exit 2` si una
de las dos carpetas no existe o no se puede abrir un fichero.
"""
import argparse
import json
import os
import sys

import openpyxl


def _celdas(ws):
    """(valores, locked) de una hoja. `valores` sólo con las celdas escritas."""
    valores, locked = {}, {}
    for row in ws.iter_rows():
        for c in row:
            locked[c.coordinate] = bool(c.protection.locked)
            if c.value is None:
                continue
            valores[c.coordinate] = (repr(c.value), c.data_type,
                                     c.number_format)
    return valores, locked


def _dv(ws):
    """Validaciones de datos de una hoja, indexadas por rango."""
    fuera = {}
    for dv in ws.data_validations.dataValidation:
        clave = str(dv.sqref)
        fuera[clave] = (dv.type, dv.formula1, dv.formula2, dv.operator,
                        bool(dv.allow_blank), bool(dv.showErrorMessage),
                        dv.errorTitle, dv.error, dv.promptTitle, dv.prompt)
    return fuera


def _alturas(ws):
    return {r: d.height for r, d in ws.row_dimensions.items()
            if d.height is not None}


def _ficheros(carpeta):
    return sorted(n for n in os.listdir(carpeta)
                  if n.endswith('.xlsx') and not n.startswith('~$'))


def comparar_libro(pa, pb, fname, difs):
    """Compara dos .xlsx y acumula las diferencias en `difs`."""
    try:
        wa = openpyxl.load_workbook(pa)
        wb = openpyxl.load_workbook(pb)
    except Exception as e:                                  # noqa: BLE001
        difs.append({'fichero': fname, 'tipo': 'apertura',
                     'detalle': f'no se pudo abrir: {e}'})
        return
    if wa.sheetnames != wb.sheetnames:
        difs.append({'fichero': fname, 'tipo': 'hojas',
                     'a': wa.sheetnames, 'b': wb.sheetnames,
                     'detalle': f'{fname}: hojas A={wa.sheetnames} '
                                f'B={wb.sheetnames}'})
    for titulo in wa.sheetnames:
        if titulo not in wb.sheetnames:
            continue
        ha, hb = wa[titulo], wb[titulo]
        va, la = _celdas(ha)
        vb, lb = _celdas(hb)
        for coord in sorted(set(va) | set(vb), key=_orden):
            if va.get(coord) != vb.get(coord):
                difs.append({
                    'fichero': fname, 'hoja': titulo, 'celda': coord,
                    'tipo': 'valor',
                    'a': va.get(coord), 'b': vb.get(coord),
                    'detalle': f'{fname}!{titulo}!{coord}: A={va.get(coord)} '
                               f'B={vb.get(coord)}'})
        for coord in sorted(set(la) | set(lb), key=_orden):
            if la.get(coord) != lb.get(coord):
                difs.append({
                    'fichero': fname, 'hoja': titulo, 'celda': coord,
                    'tipo': 'locked',
                    'a': la.get(coord), 'b': lb.get(coord),
                    'detalle': f'{fname}!{titulo}!{coord}: locked '
                               f'A={la.get(coord)} B={lb.get(coord)}'})
        da, db = _dv(ha), _dv(hb)
        for ref in sorted(set(da) | set(db)):
            if da.get(ref) != db.get(ref):
                difs.append({
                    'fichero': fname, 'hoja': titulo, 'celda': ref,
                    'tipo': 'dv', 'a': da.get(ref), 'b': db.get(ref),
                    'detalle': f'{fname}!{titulo}!{ref}: DV A={da.get(ref)} '
                               f'B={db.get(ref)}'})
        aa, ab = _alturas(ha), _alturas(hb)
        for r in sorted(set(aa) | set(ab)):
            if aa.get(r) != ab.get(r):
                difs.append({
                    'fichero': fname, 'hoja': titulo, 'celda': f'fila {r}',
                    'tipo': 'altura', 'a': aa.get(r), 'b': ab.get(r),
                    'detalle': f'{fname}!{titulo}!fila {r}: alto '
                               f'A={aa.get(r)} B={ab.get(r)}'})


def _orden(coord):
    """Orden natural de una coordenada: A5 antes que A10 y que B1."""
    letras = ''.join(ch for ch in coord if ch.isalpha())
    numero = ''.join(ch for ch in coord if ch.isdigit())
    return (letras.rjust(3), int(numero) if numero else 0)


def comparar(a, b, solo=None):
    """Compara dos carpetas. Devuelve el informe como dict."""
    for carpeta in (a, b):
        if not os.path.isdir(carpeta):
            raise SystemExit(f'ABORTADO: no existe la carpeta {carpeta}')
    fa, fb = _ficheros(a), _ficheros(b)
    if solo:
        fa = [n for n in fa if any(s in n for s in solo)]
        fb = [n for n in fb if any(s in n for s in solo)]
    difs = []
    for n in sorted(set(fa) - set(fb)):
        difs.append({'fichero': n, 'tipo': 'ausente',
                     'detalle': f'{n}: sólo está en A'})
    for n in sorted(set(fb) - set(fa)):
        difs.append({'fichero': n, 'tipo': 'ausente',
                     'detalle': f'{n}: sólo está en B'})
    comunes = sorted(set(fa) & set(fb))
    for n in comunes:
        comparar_libro(os.path.join(a, n), os.path.join(b, n), n, difs)
    por_fichero = {}
    for d in difs:
        por_fichero[d['fichero']] = por_fichero.get(d['fichero'], 0) + 1
    por_tipo = {}
    for d in difs:
        por_tipo[d['tipo']] = por_tipo.get(d['tipo'], 0) + 1
    return {'a': a, 'b': b, 'ficheros_comparados': comunes,
            'diferencias': len(difs), 'por_fichero': por_fichero,
            'por_tipo': por_tipo, 'detalle': difs}


def main():
    ap = argparse.ArgumentParser(
        description='Compara dos carpetas de .xlsx celda a celda '
                    '(valor, fórmula, DV, formato, locked, alto de fila)')
    ap.add_argument('--a', required=True, help='carpeta A (p.ej. producción)')
    ap.add_argument('--b', required=True, help='carpeta B (p.ej. el dry-run)')
    ap.add_argument('--json', default=None)
    ap.add_argument('--max', type=int, default=40,
                    help='diferencias que se imprimen (el JSON las lleva '
                         'TODAS)')
    ap.add_argument('--solo', default=None,
                    help='subcadenas de nombre de fichero, separadas por coma')
    args = ap.parse_args()
    solo = [s.strip() for s in args.solo.split(',')] if args.solo else None
    inf = comparar(args.a, args.b, solo)
    print(f"ficheros comparados: {len(inf['ficheros_comparados'])} · "
          f"diferencias: {inf['diferencias']}")
    if inf['por_tipo']:
        print('  por tipo: ' + ' · '.join(f'{k}={v}'
                                          for k, v in sorted(
                                              inf['por_tipo'].items())))
    if inf['por_fichero']:
        print('  por fichero: ' + ' · '.join(f'{k}={v}'
                                             for k, v in sorted(
                                                 inf['por_fichero'].items())))
    for d in inf['detalle'][:args.max]:
        print('  ' + d['detalle'])
    if inf['diferencias'] > args.max:
        print(f"  … y {inf['diferencias'] - args.max} más (están en el JSON)")
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(inf, fh, ensure_ascii=False, indent=1)
        print(f'informe → {args.json}')
    return 1 if inf['diferencias'] else 0


if __name__ == '__main__':
    sys.exit(main())
