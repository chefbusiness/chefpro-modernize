#!/usr/bin/env python3
"""
construir_09_catering.py — constructor de `09-cobros-facturacion-eventos.xlsx`
(kit-tareas-catering).

    from construir_09_catering import construir
    construir('/ruta/09-cobros-facturacion-eventos.xlsx', motor.CTX)

DISEÑO FIRMADO POR JOHN (2026-08-23). El 09 de catering modelaba una caja de
mostrador —fondo, cambio, Z del TPV por turno— y eso describe un negocio que no
es el suyo: una empresa de catering **no tiene mostrador**, factura por EVENTO y
cobra mayoritariamente por transferencia (anticipo del 30-50 % + saldo). El
fichero pasa a llamarse «Cobros y Facturación por Evento — Catering / Eventos»
y sustituye a `09-apertura-cierre-caja.xlsx`. El `git mv` del entregable en
`astro-site/public/dl/` y el redirect de la descarga los hace el ORQUESTADOR,
no este script; `main.py` sabe construirlo tanto si en la copia de trabajo está
el nombre viejo como si ya sólo está el nuevo.

Cinco hojas: Instrucciones · Antes del Evento · Después del Evento ·
Liquidación del Evento · Registro de Eventos.

REFERENCIA VISUAL — se ha copiado celda a celda de los dos ficheros de la
familia que ya están en producción (`dl/kit-tareas-catering/`):
`08-apertura-cierre-negocio.xlsx` y el propio `09-apertura-cierre-caja.xlsx`.
De ahí salen, y no de una paleta inventada: el título en blanco sobre `1A1A1A`
combinado en toda la fila, la banda dorada `FFD700` de la fila 2, la cabecera
`2D2D2D`, el zebrado `F5F5F5`/blanco de las columnas fijas, el verde `E8F5E9`
de todo lo editable, el borde `thin` en `E0E0E0`, Calibri, las alturas
(40 · 28 · 8 · 28 · 24) y el `freeze_panes` en la primera fila de datos. Un
fichero con otra paleta se ve a la primera dentro del kit.

CONTRATO CON EL MOTOR — los rótulos NO se escriben a mano: `ETIQ_EV_*`,
`CAB_EVENTOS`, `EV_COBRADO/EV_PENDIENTE/EV_VENCIDO` y `EV_TOLERANCIA` viven en
`motor.py` porque son lo que usan, a la vez, `fila_liquidacion` /
`fila_registro_eventos` (que deciden el PAPEL del fichero por cabecera, nunca
por su nombre) y `main.demo_liquidacion` (que localiza las celdas por rótulo
para demostrar §6 con pycel). Cambiar un rótulo aquí y no allí deja el fichero
sin papel, `CTX['f_caja']` a None y los 11 ficheros del kit sin la línea del
dinero en sus Instrucciones, sin que ningún gate lo cante.

DOS INVARIANTES QUE `main.demo_liquidacion` DA POR CIERTAS y que por eso no se
pueden mover al rediseñar la hoja de liquidación: el rótulo va en la columna
**A o B** (`motor._buscar` sólo mira esas dos) y el importe en la columna **C**
(la demo escribe `ws.cell(row=..., column=3)`). Si el formulario creciera a dos
columnas de importes, hay que tocar la demo en el mismo commit.

LO QUE HACE EL MOTOR Y AQUÍ NO SE DUPLICA: las 5 filas libres de cada checklist
(§2.2), el contador honesto (§2.1), el formato condicional de fila completada,
la protección sin contraseña, el A4/print_area/pie de impresión, la bio, la
línea de versión, la hoja «Instrucciones» entera (la reconstruye
`motor.reescribir_instrucciones` desde `motor.instrucciones_cobros`) y la
metadata. Este constructor deja los centinelas que el motor necesita —cabecera
del molde ▸, fila en blanco + fila del contador, hoja «Instrucciones» con su
línea de versión— y todo lo que el motor NO sabe inventar: el contenido.

El desplegable «✓,—,N/A» sí se escribe aquí, con `motor.DV_LISTA` y
`motor.DV_ERROR`, para que el fichero sea correcto por sí solo aunque nunca
pase por el motor; el motor lo sustituye después por uno idéntico.
"""
import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import motor

# ==========================================================================
# Paleta y tipografía — medidas sobre los ficheros de producción del kit
# ==========================================================================
NEGRO = '1A1A1A'        # fondo del título de hoja y de las filas de TOTALES
CAB_FONDO = '2D2D2D'    # fondo de las cabeceras de tabla
ORO = 'FFD700'          # bandas de sección, filas de resultado y firma
BLANCO = 'FFFFFF'
#: El rojo del VENCIDO vive en `motor.py` (`ROJO`/`ROJO_TXT`) con el resto del
#: contrato del modelo por eventos: quien lo vuelve a pintar en cada pasada es
#: `motor.cf_eventos`, no este constructor.
ROJO = motor.ROJO
ROJO_TXT = motor.ROJO_TXT

F_TITULO_HOJA = Font(name='Calibri', bold=True, size=16, color='00' + BLANCO)
F_BANDA = Font(name='Calibri', bold=True, size=10, color='00' + NEGRO)
F_CAB = Font(name='Calibri', bold=True, size=11, color='00' + BLANCO)
F_SECCION = Font(name='Calibri', bold=True, size=13, color='00' + NEGRO)
F_SUBSECCION = Font(name='Calibri', bold=True, size=11, color='00' + NEGRO)
F_TXT = Font(name='Calibri', size=10, color='00' + NEGRO)
F_TXT_B = Font(name='Calibri', bold=True, size=10, color='00' + NEGRO)
F_TOTAL = Font(name='Calibri', bold=True, size=11, color='00' + BLANCO)
F_NOTA = Font(name='Calibri', size=9, color='00666666')
F_PIE = Font(name='Calibri', size=9, color='00888888')

FILL_TITULO = PatternFill('solid', fgColor=NEGRO)
FILL_CAB = PatternFill('solid', fgColor=CAB_FONDO)
FILL_ORO = PatternFill('solid', fgColor=ORO)
FILL_VERDE = PatternFill('solid', fgColor=motor.VERDE)
FILL_GRIS = PatternFill('solid', fgColor=motor.GRIS)
FILL_BLANCO = PatternFill('solid', fgColor=BLANCO)
FILL_AMBAR = PatternFill('solid', start_color=motor.AMBAR,
                         end_color=motor.AMBAR)
FILL_ROJO = PatternFill('solid', start_color=ROJO, end_color=ROJO)

BORDE = Border(*[Side(style='thin', color='E0E0E0')] * 4)
CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
IZQ = Alignment(horizontal='left', vertical='center', wrap_text=True)
DER = Alignment(horizontal='right', vertical='center', wrap_text=True)

#: Cabecera del molde ▸ de los dos checklists. «Cuándo» y no «Hora Límite»: la
#: columna lleva hitos en días ANTES/DESPUÉS del evento (D-15, D+7) y no horas
#: de reloj — `motor.cadencia` la clasificaría igual (DOM-R2-24), así que se
#: escribe ya bien y la 1.ª pasada no tiene nada que corregir. «#» y no «Nº»:
#: es lo que exige `motor.cabecera_checklist` y lo que llevan los otros 10.
CAB_CHECK = ('#', 'Tarea', 'Responsable', 'Cuándo', motor.CAB_MARCA, 'Firma',
             'Notas')
ANCHOS_CHECK = ((10, 50, 20, 13, 14, 14, 25))

# ==========================================================================
# Contenido — brief de John (2026-08-23), redactado para el cliente final
# ==========================================================================
#: 12 tareas + las 5 filas libres del motor. Los «Cuándo» son hitos de cuenta
#: atrás (D-15 … D-1) porque en catering el calendario lo marca la FECHA DEL
#: EVENTO, no la hora de apertura de un local: un anticipo se pide al firmar,
#: no «a las 07:00».
ANTES = (
    ('Presupuesto del evento firmado o aceptado por escrito (email o PDF '
     'firmado vale; una confirmación de palabra, no)', 'Comercial', 'D-15'),
    ('Nº de comensales confirmado y fecha límite de cambios comunicada por '
     'escrito al cliente', 'Comercial', 'D-15'),
    ('Anticipo del 30-50 % cobrado y registrado: fecha, importe y medio de '
     'pago', 'Administración', 'D-15'),
    ('Forma de pago del saldo acordada (transferencia, tarjeta o efectivo) y '
     'plazo de pago por escrito', 'Administración', 'D-15'),
    ('Datos de facturación completos: razón social, CIF/NIF, dirección y '
     'email de facturación', 'Administración', 'D-15'),
    ('Comprobado que el cliente no arrastra facturas anteriores pendientes '
     '(mira el «Registro de Eventos»)', 'Administración', 'D-15'),
    ('Condiciones de cancelación y de modificación comunicadas por escrito y '
     'aceptadas', 'Comercial', 'D-7'),
    ('Proveedores externos del evento (carpa, sonido, alquiler de menaje) con '
     'el pedido cerrado y su pago o anticipo hecho', 'Event Manager', 'D-7'),
    ('Escandallo del evento cerrado y margen comprobado antes de producir',
     'Dirección', 'D-7'),
    ('Seguro de responsabilidad civil del evento en vigor (y el que exija el '
     'recinto, si lo pide)', 'Dirección', 'D-7'),
    ('Alérgenos e intolerancias recibidos por escrito del cliente y pasados a '
     'cocina', 'Event Manager', 'D-3'),
    ('Solo si habrá barra con cobro en efectivo: fondo de caja preparado y '
     'anotado en la sección opcional de «Liquidación del Evento»',
     'Administración', 'D-1'),
)
#: 12 tareas. «Cuándo» en D+0 … D+30: el día del evento se cuenta y se anota,
#: al día siguiente se factura y se comunica el saldo, a la semana se cobra y
#: se concilia, y al mes se archiva el expediente y se reclama lo vencido.
DESPUES = (
    ('Recuento de comensales REALES frente a los contratados, con el dato '
     'anotado el mismo día', 'Event Manager', 'D+0'),
    ('Extras consumidos anotados: horas extra de personal, bebidas, '
     'suplementos y roturas', 'Event Manager', 'D+0'),
    ('Solo si hubo barra con cobro en efectivo: arqueo en la sección opcional '
     'de «Liquidación del Evento»', 'Administración', 'D+0'),
    ('Cargos por roturas o incidencias comunicados al cliente antes de '
     'facturar, no dentro de la factura', 'Event Manager', 'D+1'),
    ('Factura emitida con el desglose de IVA: 10 % en alimentos y bebidas no '
     'alcohólicas del servicio de catering y 21 % en alquileres, decoración, '
     'servicios y bebidas alcohólicas, tal y como lo trate tu asesor',
     'Administración', 'D+1'),
    ('Saldo (total − anticipo) comunicado al cliente con su fecha de '
     'vencimiento', 'Administración', 'D+1'),
    ('«Liquidación del Evento» rellenada y volcada como una fila del '
     '«Registro de Eventos»', 'Administración', 'D+1'),
    ('Cobro del saldo registrado: fecha, medio de pago y referencia',
     'Administración', 'D+7'),
    ('Cobro conciliado con el extracto bancario (que el importe esté en la '
     'cuenta, no solo prometido)', 'Administración', 'D+7'),
    ('Valoración o reseña pedida al cliente mientras el evento está reciente',
     'Comercial', 'D+7'),
    ('Expediente del evento archivado: presupuesto, factura, BEO e '
     'incidencias', 'Administración', 'D+30'),
    ('Saldos en estado VENCIDO del «Registro de Eventos» revisados y '
     'reclamados uno a uno', 'Dirección', 'D+30'),
)
#: Denominaciones del recuento OPCIONAL de la barra en efectivo. Mismo juego
#: que el modelo de mostrador, con la moneda de 0,02 € que DOM-03 echó en falta.
BILLETES = (('500 €', 500), ('200 €', 200), ('100 €', 100), ('50 €', 50),
            ('20 €', 20), ('10 €', 10), ('5 €', 5))
MONEDAS = (('2 €', 2), ('1 €', 1), ('0,50 €', 0.5), ('0,20 €', 0.2),
           ('0,10 €', 0.1), ('0,05 €', 0.05), ('0,02 €', 0.02),
           ('0,01 €', 0.01))

#: Rótulos que NO viven en `motor.py` porque ningún gate ni ninguna detección
#: los mira: son sólo texto de la sección opcional y de los totales.
ETIQ_BARRA = 'Solo si hubo barra con cobro en EFECTIVO'
ETIQ_TOTAL_EF = 'TOTAL EFECTIVO'
ETIQ_FONDO_BARRA = 'Fondo de la barra (−)'
ETIQ_NETO = 'Efectivo neto del evento'
FIRMA = ('Firma del responsable: _________________________     '
         'Fecha: ____/____/________')


# ==========================================================================
# Piezas de estilo
# ==========================================================================
def _titulo(ws, texto, ncol):
    """Fila 1: título de hoja, blanco sobre negro, combinado en toda la fila."""
    cel = ws.cell(row=1, column=1, value=texto)
    cel.font, cel.fill, cel.alignment = F_TITULO_HOJA, FILL_TITULO, CENTRO
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.row_dimensions[1].height = 40


def _banda_oro(ws, fila, tramos, ncol, alto=28):
    """Fila dorada de datos del documento. `tramos` = ((col_ini, texto), …)."""
    cortes = [c for c, _ in tramos] + [ncol + 1]
    for i, (col, texto) in enumerate(tramos):
        cel = ws.cell(row=fila, column=col, value=texto)
        cel.font, cel.fill, cel.alignment = F_BANDA, FILL_ORO, IZQ
        if cortes[i + 1] - 1 > col:
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila,
                           end_column=cortes[i + 1] - 1)
    for c in range(1, ncol + 1):
        ws.cell(row=fila, column=c).fill = FILL_ORO
    ws.row_dimensions[fila].height = alto


def _separador(ws, fila):
    ws.row_dimensions[fila].height = 8


def _cabecera(ws, fila, rotulos, col0=1):
    for i, v in enumerate(rotulos):
        cel = ws.cell(row=fila, column=col0 + i, value=v)
        cel.font, cel.fill = F_CAB, FILL_CAB
        cel.alignment, cel.border = CENTRO, BORDE
    ws.row_dimensions[fila].height = 28


def _seccion(ws, fila, texto, ncol):
    """Banda dorada de sección (el «💰 Recuento de Efectivo» de la familia)."""
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font, cel.fill, cel.alignment = F_SECCION, FILL_ORO, IZQ
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                   end_column=ncol)
    ws.row_dimensions[fila].height = 26


def _subseccion(ws, fila, texto, ncol):
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font, cel.fill, cel.alignment = F_SUBSECCION, FILL_ORO, IZQ
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                   end_column=ncol)
    ws.row_dimensions[fila].height = 20


def _firma_y_pie(ws, fila, ncol):
    cel = ws.cell(row=fila, column=1, value=FIRMA)
    cel.font, cel.fill, cel.alignment = F_TXT_B, FILL_ORO, IZQ
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                   end_column=ncol)
    ws.row_dimensions[fila].height = 30
    # El pie definitivo lo reescribe `motor.cerrar` con `CTX['pie']` (que lleva
    # el nombre del kit); aquí se deja uno con la MISMA forma para que el
    # fichero sea correcto por sí solo y para que `RX_PIE` lo reconozca.
    pie = ws.cell(row=fila + 1, column=1,
                  value='— Kit de Tareas Recurrentes · Catering / Eventos · '
                        'AI Chef Pro · aichef.pro')
    pie.font, pie.alignment = F_PIE, IZQ
    ws.merge_cells(start_row=fila + 1, start_column=1, end_row=fila + 1,
                   end_column=ncol)
    return fila + 1


def _anchos(ws, anchos, col0=1):
    for i, ancho in enumerate(anchos):
        ws.column_dimensions[get_column_letter(col0 + i)].width = ancho


def _zebra(fila_datos):
    """Blanco/gris alternos, empezando en gris como en toda la familia."""
    return FILL_GRIS if fila_datos % 2 == 1 else FILL_BLANCO


# ==========================================================================
# Hojas
# ==========================================================================
def _checklist(wb, titulo, subtitulo, tareas):
    """Una hoja del molde ▸.

    Deja los DOS centinelas que `motor.normalizar_checklist` necesita: la
    cabecera en la fila 4 (que fija `hr`) y, tras una fila en blanco, la fila
    del contador. El motor mete sus 5 filas libres ENTRE la última tarea y el
    contador, así que quedan DENTRO del rango contado (§2.2) — que es lo que
    hace honesto al contador cuando el cliente añade tareas propias.
    """
    ws = wb.create_sheet(titulo)
    ncol = len(CAB_CHECK)
    _anchos(ws, ANCHOS_CHECK)
    _titulo(ws, titulo + ' — Catering / Eventos', ncol)
    _banda_oro(ws, 2, ((1, 'Evento / Cliente: _________________________'),
                       (4, subtitulo)), ncol)
    _separador(ws, 3)
    _cabecera(ws, 4, CAB_CHECK)
    dv = DataValidation(type='list', formula1=motor.DV_LISTA, allow_blank=True,
                        showErrorMessage=True, errorStyle='stop',
                        errorTitle=motor.DV_ERROR_TIT, error=motor.DV_ERROR)
    ws.add_data_validation(dv)
    fila = 5
    for i, (texto, resp, cuando) in enumerate(tareas, start=1):
        fondo = _zebra(i)
        ws.cell(row=fila, column=1, value=i)
        ws.cell(row=fila, column=2, value=texto)
        ws.cell(row=fila, column=3, value=resp)
        ws.cell(row=fila, column=4, value=cuando)
        for c in range(1, ncol + 1):
            cel = ws.cell(row=fila, column=c)
            cel.font, cel.border = F_TXT, BORDE
            cel.alignment = IZQ if c == 2 else CENTRO
            # Columnas fijas (# y Tarea) zebradas; todo lo que el operario
            # escribe o cambia, en verde: es la promesa de la landing.
            cel.fill = fondo if c <= 2 else FILL_VERDE
        dv.add(ws.cell(row=fila, column=5))
        ws.row_dimensions[fila].height = 24
        fila += 1
    fila += 1                                   # fila en blanco (centinela)
    ws.cell(row=fila, column=1, value=motor.ETIQ_CONTADOR).font = F_TXT_B
    ws.cell(row=fila, column=1).alignment = IZQ
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    ws.row_dimensions[fila].height = 22
    # El contador honesto (fórmulas de C, «de» en D y el denominador en E) lo
    # escribe `motor._contador` con el rango YA ampliado por las filas libres.
    ws.freeze_panes = 'A5'
    return _firma_y_pie(ws, fila + 2, ncol)


def _liquidacion(wb):
    """Formulario de liquidación de UN evento + la barra en efectivo opcional.

    Molde: rótulo combinado en A:B, importe en C, explicación en D. Las dos
    primeras columnas y la tercera son un contrato con `main.demo_liquidacion`
    (ver la cabecera del módulo), no una preferencia estética.
    """
    ws = wb.create_sheet('Liquidación del Evento')
    ncol = 4
    _anchos(ws, (34, 22, 18, 46))
    _titulo(ws, 'Liquidación del Evento — Catering / Eventos', ncol)
    _banda_oro(ws, 2, ((1, 'Rellena SOLO las celdas verdes'),
                       (3, 'Nº de factura: ______________')), ncol)
    _separador(ws, 3)
    _cabecera(ws, 4, ('Concepto', '', 'Importe (€)', 'Qué escribir aquí'))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    ref = {}

    def fila_dato(f, etq, fmt, nota, valor=None, verde=True, formula=None,
                  destacada=False, alto=22):
        """Una línea del formulario. Verde = la escribe el cliente."""
        cel = ws.cell(row=f, column=1, value=etq)
        cel.font = F_TXT_B if destacada else F_TXT
        cel.alignment, cel.border = IZQ, BORDE
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=2)
        ws.cell(row=f, column=2).border = BORDE
        val = ws.cell(row=f, column=3, value=formula if formula else valor)
        val.font = F_TXT_B if destacada else F_TXT
        val.alignment, val.border = DER if fmt else CENTRO, BORDE
        if fmt:
            val.number_format = fmt
        nt = ws.cell(row=f, column=4, value=nota)
        nt.font, nt.alignment, nt.border = F_NOTA, IZQ, BORDE
        fondo = FILL_ORO if destacada else (FILL_VERDE if verde
                                            else _zebra(f))
        cel.fill = ws.cell(row=f, column=2).fill = (
            FILL_ORO if destacada else _zebra(f))
        val.fill = fondo
        nt.fill = _zebra(f)
        ws.row_dimensions[f].height = alto
        ref[etq] = f
        return f + 1

    f = 5
    f = fila_dato(f, motor.ETIQ_EV_EVENTO, None,
                  'Nombre del evento y del cliente que factura.')
    f = fila_dato(f, motor.ETIQ_EV_FECHA, 'DD/MM/YYYY',
                  'El día del servicio, no el de la firma del presupuesto.')
    f = fila_dato(f, motor.ETIQ_EV_PAX_CONTR, motor.FMT_ENT,
                  'Los comensales que figuran en el presupuesto aceptado.')
    f = fila_dato(f, motor.ETIQ_EV_PAX_REAL, motor.FMT_ENT,
                  'Los que hubo de verdad. La diferencia es lo que se '
                  'discute al facturar: anótala el mismo día.')

    _seccion(ws, f, '🧾 Base imponible e IVA', ncol)
    f += 1
    f = fila_dato(f, motor.ETIQ_EV_PRESUPUESTO, motor.FMT_EUR,
                  'El importe aceptado SIN IVA.', valor=0)
    f = fila_dato(f, motor.ETIQ_EV_EXTRAS, motor.FMT_EUR,
                  'Horas extra de personal, bebidas de más, suplementos y '
                  'cargos por roturas. Sin IVA.', valor=0)
    f = fila_dato(f, motor.ETIQ_EV_BASE10, motor.FMT_EUR,
                  'Alimentos y bebidas no alcohólicas del servicio de '
                  'catering.', valor=0)
    f = fila_dato(f, motor.ETIQ_EV_BASE21, motor.FMT_EUR,
                  'Alquileres, decoración, servicios y bebidas alcohólicas.',
                  valor=0)
    # El aviso INFORMA, no bloquea. El reparto entre el 10 % y el 21 % lo
    # decide el cliente con su asesor: una validación dura le impediría
    # guardar un caso legítimo (un evento 100 % al 21 %, por ejemplo). Y con
    # la hoja en blanco calla, en vez de felicitar por un cuadre de ceros.
    f = fila_dato(
        f, motor.ETIQ_EV_AVISO_BASE, None,
        'Si no cuadra, revisa el reparto: el total facturado sale de estas '
        'dos bases, no del presupuesto.', verde=False,
        formula='=IF(C{p}+C{e}+C{a}+C{b}=0,"",IF(ABS(C{a}+C{b}-C{p}-C{e})<={t},'
                '"OK: las dos bases suman presupuesto + extras",'
                '"REVISA: las dos bases no suman presupuesto + extras"))'
                .format(p=ref[motor.ETIQ_EV_PRESUPUESTO],
                        e=ref[motor.ETIQ_EV_EXTRAS],
                        a=ref[motor.ETIQ_EV_BASE10],
                        b=ref[motor.ETIQ_EV_BASE21], t=motor.EV_TOLERANCIA),
        alto=30)
    ws.cell(row=ref[motor.ETIQ_EV_AVISO_BASE], column=3).alignment = CENTRO
    f = fila_dato(f, motor.ETIQ_EV_IVA10, motor.FMT_EUR, 'Fórmula.',
                  verde=False,
                  formula='=IFERROR(ROUND(C{}*0.1,2),0)'
                          .format(ref[motor.ETIQ_EV_BASE10]))
    f = fila_dato(f, motor.ETIQ_EV_IVA21, motor.FMT_EUR, 'Fórmula.',
                  verde=False,
                  formula='=IFERROR(ROUND(C{}*0.21,2),0)'
                          .format(ref[motor.ETIQ_EV_BASE21]))
    f = fila_dato(f, motor.ETIQ_EV_TOTAL, motor.FMT_EUR,
                  'Las dos bases más sus dos IVA. Es lo que dice la factura.',
                  verde=False, destacada=True,
                  formula='=IFERROR(C{}+C{}+C{}+C{},0)'.format(
                      ref[motor.ETIQ_EV_BASE10], ref[motor.ETIQ_EV_BASE21],
                      ref[motor.ETIQ_EV_IVA10], ref[motor.ETIQ_EV_IVA21]))

    _seccion(ws, f, '💶 Anticipo, saldo y cobro', ncol)
    f += 1
    f = fila_dato(f, motor.ETIQ_EV_ANTICIPO, motor.FMT_EUR,
                  'Lo cobrado al firmar (el 30-50 % habitual). Se resta del '
                  'total.', valor=0)
    f = fila_dato(f, motor.ETIQ_EV_SALDO, motor.FMT_EUR,
                  'Fórmula: total − anticipo. Es lo que se factura al final.',
                  verde=False,
                  formula='=IFERROR(C{}-C{},0)'.format(
                      ref[motor.ETIQ_EV_TOTAL], ref[motor.ETIQ_EV_ANTICIPO]))
    f = fila_dato(f, motor.ETIQ_EV_COBRADO, motor.FMT_EUR,
                  'Lo que ha entrado DESPUÉS del evento y está conciliado con '
                  'el banco.', valor=0)
    f = fila_dato(f, motor.ETIQ_EV_PENDIENTE, motor.FMT_EUR,
                  'Fórmula: saldo − cobrado. En ámbar mientras quede algo por '
                  'cobrar.', verde=False, destacada=True,
                  formula='=IFERROR(C{}-C{},0)'.format(
                      ref[motor.ETIQ_EV_SALDO], ref[motor.ETIQ_EV_COBRADO]))
    f = fila_dato(f, motor.ETIQ_EV_VENCIMIENTO, 'DD/MM/YYYY',
                  'La fecha que acordaste con el cliente. Sin ella el estado '
                  'nunca puede decir VENCIDO.')
    f = fila_dato(f, motor.ETIQ_EV_ESTADO, None,
                  '«{}» si no queda ni un céntimo · «{}» si aún hay saldo · '
                  '«{}» si además ya pasó el vencimiento.'
                  .format(motor.EV_COBRADO, motor.EV_PENDIENTE,
                          motor.EV_VENCIDO),
                  verde=False, destacada=True,
                  formula='=IF(C{tot}<=0,"",IF(C{p}<={t},"{ok}",'
                          'IF(C{v}="","{pe}",IF(TODAY()>C{v},"{ve}","{pe}"))))'
                          .format(tot=ref[motor.ETIQ_EV_TOTAL],
                                  p=ref[motor.ETIQ_EV_PENDIENTE],
                                  t=motor.EV_TOLERANCIA,
                                  v=ref[motor.ETIQ_EV_VENCIMIENTO],
                                  ok=motor.EV_COBRADO, pe=motor.EV_PENDIENTE,
                                  ve=motor.EV_VENCIDO), alto=26)
    # Ámbar si queda PENDIENTE y rojo si el saldo está VENCIDO. Lo pinta
    # `motor.cf_eventos` y no este fichero: `motor.aplicar` VACÍA el formato
    # condicional de toda hoja reconocida para ser idempotente, así que un CF
    # escrito sólo aquí existe en el fichero recién construido y desaparece en
    # la primera pasada del motor — con el dry-run en verde. Se llama igual
    # para que el constructor suelto ya entregue el fichero completo.
    motor.cf_eventos(ws, 'liquidacion')

    # --- sección OPCIONAL: barra con cobro en efectivo --------------------
    f += 1
    _seccion(ws, f, '💰 ' + ETIQ_BARRA, ncol)
    f += 1
    nota = ws.cell(row=f, column=1,
                   value='Cuenta el efectivo y, si quieres, suma tú el '
                         '«{}» a «{}». NO se enlaza solo: en la mayoría de '
                         'los eventos no hay barra en efectivo y una fórmula '
                         'fija dejaría un 0 restando donde no debe.'
                         .format(ETIQ_NETO, motor.ETIQ_EV_COBRADO))
    nota.font, nota.alignment = F_NOTA, IZQ
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=ncol)
    ws.row_dimensions[f].height = 30
    f += 1
    _cabecera(ws, f, ('Denominación', 'Cantidad', 'Subtotal (€)',
                      'Anota sólo lo que haya en el cajón'))
    f += 1
    primera = None
    for banda, filas in (('BILLETES', BILLETES), ('MONEDAS', MONEDAS)):
        _subseccion(ws, f, banda, ncol)
        f += 1
        for i, (etq, valor) in enumerate(filas):
            if primera is None:
                primera = f
            fondo = _zebra(i)
            a = ws.cell(row=f, column=1, value=etq)
            a.font, a.fill, a.alignment, a.border = F_TXT_B, fondo, IZQ, BORDE
            b = ws.cell(row=f, column=2, value=0)
            b.font, b.fill, b.alignment, b.border = (F_TXT, FILL_VERDE, DER,
                                                     BORDE)
            b.number_format = motor.FMT_ENT
            c = ws.cell(row=f, column=3,
                        value='=IFERROR(B{}*{},0)'.format(f, valor))
            c.font, c.fill, c.alignment, c.border = F_TXT, fondo, DER, BORDE
            c.number_format = motor.FMT_EUR
            ws.cell(row=f, column=4).fill = fondo
            ws.cell(row=f, column=4).border = BORDE
            ws.row_dimensions[f].height = 20
            f += 1
    ultima = f - 1
    tot = ws.cell(row=f, column=1, value=ETIQ_TOTAL_EF)
    tot.font, tot.fill, tot.alignment = F_TOTAL, FILL_TITULO, IZQ
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=2)
    ws.cell(row=f, column=2).fill = FILL_TITULO
    c = ws.cell(row=f, column=3,
                value='=IFERROR(SUM(C{}:C{}),0)'.format(primera, ultima))
    c.font, c.fill, c.alignment = F_TOTAL, FILL_TITULO, DER
    c.number_format = motor.FMT_EUR
    ws.cell(row=f, column=4).fill = FILL_TITULO
    ws.row_dimensions[f].height = 24
    r_total_ef = f
    f += 1
    f = fila_dato(f, ETIQ_FONDO_BARRA, motor.FMT_EUR,
                  'El cambio con el que salió la barra. No es venta.',
                  valor=0)
    f = fila_dato(f, ETIQ_NETO, motor.FMT_EUR,
                  'Fórmula: total contado − fondo. Esto sí es cobro del '
                  'evento.', verde=False, destacada=True,
                  formula='=IFERROR(C{}-C{},0)'.format(
                      r_total_ef, ref[ETIQ_FONDO_BARRA]))
    ws.freeze_panes = 'A5'
    return _firma_y_pie(ws, f + 1, ncol)


def _registro(wb):
    """Una fila por evento: 25 filas + TOTALES + los dos recuentos.

    La cabecera es lo que `motor.fila_registro_eventos` detecta por PREFIJO
    (`motor.CAB_EVENTOS`), así que los seis rótulos de esa constante tienen que
    seguir empezando estas columnas aunque se les cambie la cola.
    """
    ws = wb.create_sheet('Registro de Eventos')
    rot = ('Fecha', motor.ETIQ_EV_EVENTO, 'Comensales',
           'Base (presupuesto + extras)', 'Total factura', 'Anticipo',
           'Cobrado', 'Pendiente', 'Medio de pago', 'Vencimiento', 'Estado')
    ncol = len(rot)
    _anchos(ws, (12, 32, 12, 22, 16, 14, 14, 14, 18, 14, 14))
    _titulo(ws, 'Registro de Eventos — Catering / Eventos', ncol)
    _banda_oro(ws, 2, ((1, 'Año: ________'),
                       (7, 'Responsable de facturación: ______________')),
               ncol)
    # IVA medio: el registro es la vista de AÑO y no reparte base 10/21 evento
    # a evento — eso lo hace la hoja de liquidación. Se ha elegido la FÓRMULA
    # (y no una celda verde que se copie de la liquidación) porque un total
    # tecleado a mano es justo lo que hace que las dos hojas dejen de cuadrar:
    # así el registro siempre es coherente consigo mismo y quien necesite el
    # importe exacto de una factura lo tiene en «Liquidación del Evento».
    et = ws.cell(row=3, column=1, value='IVA medio aplicado a «Total factura»')
    et.font, et.alignment = F_TXT, IZQ
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    iva = ws.cell(row=3, column=3, value=0.1)
    iva.font, iva.fill, iva.alignment, iva.border = (F_TXT, FILL_VERDE, CENTRO,
                                                     BORDE)
    iva.number_format = '0 %'
    nt = ws.cell(row=3, column=4,
                 value='Cámbialo si tus eventos no son mayoritariamente al '
                       '10 %. El desglose exacto de cada factura está en '
                       '«Liquidación del Evento».')
    nt.font, nt.alignment = F_NOTA, IZQ
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=ncol)
    ws.row_dimensions[3].height = 26
    _separador(ws, 4)
    hr = 5
    _cabecera(ws, hr, rot)
    primera, ultima = hr + 1, hr + 25
    for r in range(primera, ultima + 1):
        fondo = _zebra(r - primera)
        for c in range(1, ncol + 1):
            cel = ws.cell(row=r, column=c)
            cel.font, cel.border, cel.alignment = F_TXT, BORDE, CENTRO
            cel.fill = fondo
        for c in (1, 2, 3, 4, 6, 7, 9, 10):          # lo que escribe el cliente
            ws.cell(row=r, column=c).fill = FILL_VERDE
        ws.cell(row=r, column=2).alignment = IZQ
        ws.cell(row=r, column=1).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=10).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=3).number_format = motor.FMT_ENT
        for c in (4, 6, 7):
            ws.cell(row=r, column=c).value = 0
            ws.cell(row=r, column=c).number_format = motor.FMT_EUR
            ws.cell(row=r, column=c).alignment = DER
        e = ws.cell(row=r, column=5,
                    value='=IFERROR(IF(D{r}=0,0,ROUND(D{r}*(1+$C$3),2)),0)'
                          .format(r=r))
        e.number_format, e.alignment = motor.FMT_EUR, DER
        h = ws.cell(row=r, column=8,
                    value='=IFERROR(IF(E{r}=0,0,E{r}-F{r}-G{r}),0)'.format(r=r))
        h.number_format, h.alignment = motor.FMT_EUR, DER
        ws.cell(row=r, column=11).value = (
            '=IF(E{r}=0,"",IF(H{r}<={t},"{ok}",IF(J{r}="","{pe}",'
            'IF(TODAY()>J{r},"{ve}","{pe}"))))'.format(
                r=r, t=motor.EV_TOLERANCIA, ok=motor.EV_COBRADO,
                pe=motor.EV_PENDIENTE, ve=motor.EV_VENCIDO))
        ws.row_dimensions[r].height = 22
    f = ultima + 1
    tot = ws.cell(row=f, column=1, value='TOTALES')
    tot.font, tot.fill, tot.alignment = F_TOTAL, FILL_TITULO, IZQ
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
    for c in range(1, ncol + 1):
        cel = ws.cell(row=f, column=c)
        cel.fill = FILL_TITULO
        if c in (4, 5, 6, 7, 8):
            letra = get_column_letter(c)
            cel.value = '=IFERROR(SUM({l}{a}:{l}{b}),0)'.format(
                l=letra, a=primera, b=ultima)
            cel.number_format = motor.FMT_EUR
        cel.font, cel.alignment = F_TOTAL, DER
    ws.cell(row=f, column=1).alignment = IZQ
    ws.row_dimensions[f].height = 28
    f += 2
    for etiqueta, estado, nota in (
            ('Eventos PENDIENTES de cobro', motor.EV_PENDIENTE,
             'Con saldo por cobrar y todavía en plazo.'),
            ('Eventos VENCIDOS', motor.EV_VENCIDO,
             'Pasado el vencimiento y sin cobrar: reclámalos hoy.')):
        cel = ws.cell(row=f, column=1, value=etiqueta)
        cel.font, cel.alignment = F_TXT_B, IZQ
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
        n = ws.cell(row=f, column=4,
                    value='=COUNTIF(K{}:K{},"{}")'.format(primera, ultima,
                                                         estado))
        n.number_format, n.font, n.alignment, n.border = (motor.FMT_ENT,
                                                          F_TXT_B, CENTRO,
                                                          BORDE)
        n.fill = FILL_ORO if estado == motor.EV_VENCIDO else FILL_GRIS
        c = ws.cell(row=f, column=5, value=nota)
        c.font, c.alignment = F_NOTA, IZQ
        ws.merge_cells(start_row=f, start_column=5, end_row=f, end_column=ncol)
        ws.row_dimensions[f].height = 22
        f += 1
    # Después de TOTALES: `motor.cf_eventos` delimita el cuerpo por esa fila.
    motor.cf_eventos(ws, 'registro_eventos')
    ws.freeze_panes = 'A6'
    return _firma_y_pie(ws, f + 1, ncol)


# ==========================================================================
def construir(ruta_destino, ctx=None):
    """Genera el fichero completo. `ctx` = `motor.CTX` (para la metadata)."""
    ctx = ctx or {}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Instrucciones'
    # Ancla mínima: `motor.reescribir_instrucciones` reconstruye esta hoja
    # entera desde `motor.instrucciones_cobros()`. Lo único que tiene que
    # encontrar es la hoja, su título y una línea de versión donde anclar la
    # bio (§2.6) — pero se deja íntegra y legible para que el fichero valga
    # aunque nunca pase por el motor.
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = motor.ANCHO_B
    ws.cell(row=2, column=2,
            value='Cobros y Facturación por Evento — Catering / Eventos'
            ).font = motor.F_TITULO
    ws.cell(row=4, column=2, value='Qué resuelve').font = motor.F_CAB
    ws.cell(row=6, column=2,
            value='▸ El DINERO de cada evento: el anticipo que se cobra al '
                  'firmar, la factura con su desglose de IVA, el saldo que '
                  'queda y la fecha en que vence. No es una caja de '
                  'mostrador: aquí no hay turno ni Z del TPV.'
            ).font = motor.F_TXT
    ws.cell(row=8, column=2, value=motor.BIO).font = motor.F_CAB
    ws.cell(row=10, column=2, value=motor.version_line()).font = motor.F_CAB

    _checklist(wb, 'Antes del Evento',
               'Fecha del evento: ____/____/________    '
               'Responsable: _________________________', ANTES)
    _checklist(wb, 'Después del Evento',
               'Fecha del evento: ____/____/________    '
               'Responsable: _________________________', DESPUES)
    _liquidacion(wb)
    _registro(wb)

    # Metadata en la forma de la familia (m5). `main.procesar` vuelve a
    # llamar a `motor.set_metadata` después de `cerrar`, así que esto es la
    # red de seguridad para cuando el constructor se usa suelto.
    sufijo = ctx.get('sufijo') or 'Kit de Tareas Recurrentes Pro'
    kit = ctx.get('kit') or 'Catering / Eventos'
    p = wb.properties
    p.title = 'Cobros y Facturación por Evento — {} · {}'.format(kit, sufijo)
    p.subject = sufijo + ' · v2.0'
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'
    p.keywords = (ctx.get('producto') or 'kit-tareas-catering').replace(
        '-', ' ') + ', ' + motor.COLA_KEYWORDS
    p.created = p.modified = datetime.datetime(2026, 8, 23)
    wb.save(ruta_destino)
    return ruta_destino


#: Nombres del fichero del dinero de catering. `main.py` los usa para el paso
#: de sustitución; viven aquí para que sólo haya un sitio donde cambiarlos.
NOMBRE_NUEVO = '09-cobros-facturacion-eventos.xlsx'
NOMBRE_VIEJO = '09-apertura-cierre-caja.xlsx'
PRODUCTO = 'kit-tareas-catering'


if __name__ == '__main__':
    import sys
    print(construir(sys.argv[1] if len(sys.argv) > 1 else NOMBRE_NUEVO))
