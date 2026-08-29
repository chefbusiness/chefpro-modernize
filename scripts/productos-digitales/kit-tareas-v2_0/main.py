#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 de la familia «Kit de Tareas».

    python3 main.py --producto <pid> [--dry-run] [--solo motor,contenido]
                    [--json informe.json]

`--producto` por defecto `kit-tareas` (el representante). El motor (§1 y §2 de
`kit-tareas-v2-SPEC.md`) es de FAMILIA: se puede pasar a cualquier hermano
(`kit-tareas-cafeteria`, `kit-tareas-hotel`…) y sólo tocará las hojas que
reconozca por CABECERA. El módulo de CONTENIDO es propio de cada producto y se
carga si existe `contenido_<pid con guiones bajos>.py`.

REGLA DURA: `astro-site/public/dl/<pid>/` NO se toca. `--dry-run` copia el
producto al scratchpad y trabaja allí. Sin `--dry-run` el script ABORTA salvo
que se le pase `KIT_TAREAS_APPLY=1`: la ejecución real la hace el orquestador
cuando la ronda 2 dé verde.

Orden de trabajo:
  1. Copia de trabajo (dry-run) o respaldo con marca de tiempo EN EL SCRATCHPAD
     (nunca una carpeta .bak dentro de public/dl/, que se publicaría).
  2. `motor.contexto()` lee el kit entero: nombre comercial, pie, sufijo de
     metadata, hora ancla de apertura y qué fichero hace de negocio / caja /
     áreas. De ahí sale el «Se conecta con» sin nombres a mano.
  3. Por fichero: motor.aplicar → contenido.post (si hay) → motor.cerrar.
  4. Idempotencia: se repite el pipeline sobre un clon y se compara celda a
     celda. 0 diferencias o falla.
  5. `inject_cache.py` SIEMPRE al final (cualquier guardado posterior de
     openpyxl borraría el cache).
  6. Verificación: data_only de cada fórmula nueva · pycel con los casos de §6
     (arqueo, contador honesto, denominador del 07) · DV «✓,—,N/A» en todas las
     hojas de checklist · bio en todas las Instrucciones · censo --fail.

Térmica: todo en SERIE, sin builds ni navegador.
"""
import argparse
import contextlib
import copy
import datetime
import importlib
import json
import logging
import os
import re
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    'f83a04d3-d14f-49b3-aa79-c442fb4d7983/scratchpad')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')
SPEC = 'scripts/productos-digitales/kit-tareas-v2-SPEC.md §1 y §2'


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Copias
# ==========================================================================
def preparar_copia(origen, destino):
    if not os.path.isdir(origen):
        raise SystemExit(f'No existe el origen: {origen}')
    if os.path.isdir(destino):
        shutil.rmtree(destino)
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    shutil.copytree(origen, destino)
    log(f'  copia de trabajo regenerada: {destino}')


def _orden_coord(coord):
    """Clave de orden natural de una coordenada: A5 antes que A10 y que B1."""
    m = re.match(r'^([A-Z]+)(\d+)$', coord)
    if not m:
        return (0, coord, 0)
    return (1, m.group(1).rjust(3), int(m.group(2)))


#: m5 — clave del pseudo-«hoja» que lleva las PROPIEDADES del documento en la
#: huella. Sin ellas el digest no podía demostrar ni la idempotencia de la
#: metadata ni el diff contra producción de un cambio que es SÓLO de metadata
#: (los 9 ficheros P4/BONUS de heladería): el fichero salía con «0 diferencias»
#: mientras su `subject` cambiaba de v1.1 a v2.0. Va con «·» a los lados para
#: que no pueda chocar con el título de una hoja real de Excel.
CLAVE_PROPS = '·propiedades·'
#: `lastModifiedBy`, `created` y `modified` quedan fuera a propósito: cambian en
#: cada guardado y convertirían cualquier diff en ruido.
CAMPOS_PROPS = ('title', 'subject', 'creator', 'keywords', 'category',
                'description')


def digest(path):
    """Huella comparable (valores, formato, relleno, bloqueo, merges, DV, CF).

    m3 (tanda 5) — dos puntos ciegos que hacían que el diff contra producción
    saliera CORTO, y que ya costaron una discrepancia de recuento (127 según el
    digest frente a las 177 diferencias reales del verificador):

    · **El mensaje de la validación no se comparaba.** La huella era
      `type:formula1:sqref`, así que el cambio de `DV_ERROR` en las 33 hojas de
      checklist del representante —texto que el cliente ve al equivocarse de
      marca— no aparecía por ningún lado y hubo que verificarlo a mano. Ahora
      entran `errorTitle`, `error` y `prompt`.
    · **El bloqueo sólo se miraba en las celdas CON valor.** `proteger` decide
      `locked` celda a celda, también en las vacías, y en BONUS-02 del
      representante la fila 27 está vacía: el digest veía 88 celdas
      desbloqueadas donde el fichero tiene 92. Ahora `locked` es un mapa
      aparte, de TODAS las celdas del rango usado, y sale del valor de la celda
      para que un cambio de bloqueo no se disfrace de cambio de valor.
    """
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas, locked = {}, {}
        for row in ws.iter_rows():
            for c in row:
                locked[c.coordinate] = bool(c.protection.locked)
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno)
        fuera[ws.title] = {
            'celdas': celdas,
            'locked': locked,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted('{}:{}:{}:{}:{}:{}'.format(
                dv.type, dv.formula1, dv.sqref, dv.errorTitle, dv.error,
                dv.prompt)
                for dv in ws.data_validations.dataValidation),
            'cf': sorted(str(r.sqref) for r in ws.conditional_formatting),
            'area': str(ws.print_area),
            'prot': bool(ws.protection.sheet),
            'alturas': {k: v.height for k, v in ws.row_dimensions.items()
                        if v.height},
        }
    fuera[CLAVE_PROPS] = {'props': {k: getattr(wb.properties, k, None)
                                    for k in CAMPOS_PROPS}}
    return fuera


def diff_digest(a, b, fichero):
    """Diferencias entre dos huellas, UNA LÍNEA POR COSA que cambia.

    m3 — `alturas` y `locked` se comparan elemento a elemento y no como un
    bloque: un `dict != dict` daba UNA línea por hoja con el detalle recortado a
    300 caracteres, así que ni se podían contar las categorías ni se sabía qué
    fila o qué celda había cambiado. Con el desglose, el recuento del diff
    contra producción es directamente auditable.
    """
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append(f'{fichero}:{hoja}: hoja sólo en una pasada')
            continue
        if hoja == CLAVE_PROPS:                                       # m5
            pa, pb = a[hoja]['props'], b[hoja]['props']
            for k in sorted(set(pa) | set(pb)):
                if pa.get(k) != pb.get(k):
                    fuera.append(f'{fichero}:propiedades!{k}: '
                                 f'{pa.get(k)!r} → {pb.get(k)!r}')
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'area', 'prot'):
            if ha[k] != hb[k]:
                fuera.append(f'{fichero}:{hoja}: cambia {k} '
                             f'({ha[k]} → {hb[k]})'[:300])
        aa, ab = ha['alturas'], hb['alturas']
        for f in sorted(set(aa) | set(ab)):
            if aa.get(f) != ab.get(f):
                fuera.append(f'{fichero}:{hoja}!fila {f}: altura '
                             f'{aa.get(f)} → {ab.get(f)}')
        la, lb = ha['locked'], hb['locked']
        for coord in sorted(set(la) | set(lb), key=_orden_coord):
            if la.get(coord) != lb.get(coord):
                fuera.append(f'{fichero}:{hoja}!{coord}: locked '
                             f'{la.get(coord)} → {lb.get(coord)}')
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb), key=_orden_coord):
            if ca.get(coord) != cb.get(coord):
                fuera.append(f'{fichero}:{hoja}!{coord}: '
                             f'{ca.get(coord)} → {cb.get(coord)}')
    return fuera


# ==========================================================================
# Pipeline
# ==========================================================================
def ficheros_de(carpeta):
    return sorted(f for f in os.listdir(carpeta) if f.endswith('.xlsx'))


# --------------------------------------------------------------------------
# m6/m8 — sustitución del 09 de catering
# --------------------------------------------------------------------------
def sustituir_09_catering(carpeta, pid):
    """Reconstruye el fichero del dinero de catering ANTES del contexto.

    Diseño firmado por John (2026-08-23): una empresa de catering no tiene
    mostrador, así que `09-apertura-cierre-caja.xlsx` deja de existir y en su
    lugar va `09-cobros-facturacion-eventos.xlsx`, que factura por EVENTO.

    Este paso corre sobre la COPIA DE TRABAJO y hace dos cosas, en este orden:

      1. Construye el 09 nuevo desde cero con `construir_09_catering`. Se
         construye SIEMPRE, esté o no ya el nombre nuevo en la carpeta, y aquí
         está la sutileza que hace que el paso sea idempotente **y** correcto
         después del `git mv` del orquestador: `git mv` sólo RENOMBRA, no
         cambia el contenido, así que un fichero que ya se llame
         `09-cobros-facturacion-eventos.xlsx` puede seguir siendo por dentro el
         arqueo de cajón de la v1.1. Si el paso se saltara «porque el nombre
         nuevo ya está», el kit se publicaría con el nombre bueno y el fichero
         viejo dentro, y `papel_del_fichero` lo etiquetaría como 'caja'.
      2. Borra `09-apertura-cierre-caja.xlsx` de la copia si sigue ahí. En
         producción el `git mv` lo hace el orquestador —este script no toca
         `git`—, pero `main.py` tiene que funcionar en los dos estados del
         repositorio: antes del `git mv` (está el viejo) y después (está sólo
         el nuevo).

    Corre ANTES de `motor.contexto`, que es lo que lee la carpeta para decidir
    papeles y `modelo_caja`: si corriera después, el contexto se habría
    calculado sobre el arqueo de mostrador.
    """
    import construir_09_catering as c09
    if pid != c09.PRODUCTO:
        return None
    nuevo = os.path.join(carpeta, c09.NOMBRE_NUEVO)
    viejo = os.path.join(carpeta, c09.NOMBRE_VIEJO)
    habia_viejo = os.path.isfile(viejo)
    habia_nuevo = os.path.isfile(nuevo)
    if not (habia_viejo or habia_nuevo):
        # Ni el uno ni el otro: el kit no trae fichero del dinero. No se
        # inventa uno en silencio — se avisa y decide el orquestador.
        log('  ⚠ catering: no está ni {} ni {}; no se construye nada'
            .format(c09.NOMBRE_VIEJO, c09.NOMBRE_NUEVO))
        return None
    # NO se le pasa `motor.CTX`: este paso corre ANTES de `motor.contexto`, así
    # que el global todavía trae los valores por defecto (o los del kit
    # anterior si alguien encadenara dos productos en el mismo proceso). Lo
    # único que necesita el constructor es el identificador del producto; el
    # `subject`, el `title` y las `keywords` definitivos los pone
    # `motor.set_metadata` desde `procesar`, ya con el contexto real.
    # Pero el nombre del kit y el sufijo SÍ se le pasan, leídos del título de
    # un hermano real de la misma carpeta (el patrón «<algo> — <kit> · <sufijo>»
    # que el propio motor compone): sin ellos el constructor caía al fallback
    # del representante y, como el 09 nuevo vota en `contexto()`, un empate
    # 1-1 con el 08 se desempataba alfabéticamente a favor de «… Pro» y los
    # 11 ficheros perdían el nombre del kit (cazado por el crítico-5).
    extra = {'producto': pid}
    for f in sorted(os.listdir(carpeta)):
        if not f.endswith('.xlsx') or f in (c09.NOMBRE_NUEVO, c09.NOMBRE_VIEJO):
            continue
        tit = openpyxl.load_workbook(os.path.join(carpeta, f),
                                     read_only=True).properties.title or ''
        if ' · ' in tit:
            cabeza, sufijo = tit.rsplit(' · ', 1)
            extra['sufijo'] = sufijo.strip()
            if ' — ' in cabeza:
                extra['kit'] = cabeza.rsplit(' — ', 1)[1].strip()
            break
    if 'sufijo' not in extra:
        log('  ⚠ catering: ningún hermano con título «… · <sufijo>»; el 09 '
            'se construye con el sufijo por defecto')
    c09.construir(nuevo, extra)
    if habia_viejo:
        os.remove(viejo)
    return {'construido': c09.NOMBRE_NUEVO,
            'borrado': c09.NOMBRE_VIEJO if habia_viejo else None,
            'estado_previo': ('09 viejo' if habia_viejo
                              else '09 nuevo (reconstruido)')}


def procesar(carpeta, fname, etapas, contenido, informe):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    cambios = []
    estado = None
    orto = None
    # El registro de fórmulas se vacía AQUÍ y no dentro de `motor.aplicar`:
    # aplicar() sale antes de tocarlo cuando el fichero está fuera de alcance
    # (los 01-17 del hotel), y entonces este fichero heredaba el registro del
    # anterior — 98 «fórmulas sin cache» que en realidad estaban en otro libro.
    motor.REGISTRO.clear()
    if 'motor' in etapas:
        estado = motor.aplicar(wb, fname, cambios)
    if 'contenido' in etapas and contenido is not None and \
            hasattr(contenido, 'post'):
        # El módulo de contenido devuelve True cuando ha cambiado la
        # ESTRUCTURA del libro (filas u hojas nuevas). En ese caso hay que
        # volver a pasar el motor ANTES de cerrar: `estado` se midió sobre la
        # geometría vieja y con ella (a) el cuerpo que `proteger` desbloquea se
        # quedaría corto —las últimas filas de la tabla saldrían bloqueadas—,
        # (b) el rango del CF no cubriría las filas nuevas y (c) una hoja
        # creada por el contenido no estaría en `recon`, así que se publicaría
        # sin DV, sin contador, sin A4 y sin protección. El motor es
        # idempotente, así que repasarlo no cuesta más que tiempo.
        if contenido.post(wb, fname, cambios) and 'motor' in etapas and estado:
            motor.REGISTRO.clear()
            extra = []
            estado = motor.aplicar(wb, fname, extra)
            cambios += [c for c in extra if c not in cambios]
    if 'motor' in etapas and estado:
        motor.cerrar(wb, fname, estado, cambios)
    # m5 — la metadata se fija para TODOS los ficheros del producto, estén o no
    # en alcance del molde ▸. Vivía dentro de `motor.cerrar`, que sale antes
    # cuando `estado` es None, y por eso los ficheros del molde P4 y los dos
    # BONUS de los cinco kits con alcance «sólo 08/09» se publicaban con
    # `subject` «… · v1.1» dentro de un producto v2.0 — guardados (reciben
    # desplegable, contador y bio) pero sin tocar sus propiedades. Va DESPUÉS de
    # `cerrar` porque el título sale de la hoja «Instrucciones», que `cerrar`
    # reescribe, y ANTES de calcular `guardado` para que un fichero cuyo único
    # cambio sea la metadata también se guarde.
    if 'motor' in etapas:
        motor.set_metadata(wb, fname, cambios)
        # CB-E1 — la ortografía va AQUÍ, la última: `motor.cerrar` reconstruye
        # «Instrucciones» leyendo el texto que ya había y reinyecta el que
        # venía sin tildes (demostrado: tras el dry-run del 2026-08-29,
        # `01:Instrucciones!B4` seguía diciendo «Como usar estas plantillas»).
        # Corriendo antes, el motor lo desharía (§7-bis.11). Va también DESPUÉS
        # de `set_metadata`, que compone el `title` desde esa misma hoja, y
        # ANTES de `inject_cache`, que es lo último de todo el pipeline.
        orto = motor.ortografia(wb, fname, cambios)
    registro = list(motor.REGISTRO)
    guardado = bool(estado or cambios)
    if guardado:
        # Guardar con openpyxl BORRA el valor cacheado de TODAS las fórmulas del
        # libro, también de las que el motor no ha tocado (y de las hojas fuera
        # de alcance, como los inventarios del 05 de cafetería). Por eso
        # `inject_cache` se corre sobre todo fichero GUARDADO, no sólo sobre los
        # que estrenan fórmula: si no, el censo saca `nocache_real` en ficheros
        # que el motor apenas rozó.
        wb.save(path)
    informe.append({'fichero': fname,
                    'guardado': guardado,
                    'en_alcance': bool(estado),
                    'hojas_tocadas': sorted((estado or {}).get('hojas', {})),
                    'hojas_por_tipo': (estado or {}).get('hojas', {}),
                    'cambios': cambios,
                    'ortografia': orto,
                    'formulas_nuevas': len(registro)})
    return registro


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        fuera.append({'fichero': n, 'salida': r.stdout.strip(),
                      'exit': r.returncode})
        log('    ' + (r.stdout.strip() or r.stderr.strip()[-200:]))
    return fuera


def verificar_cache(carpeta, registros):
    fallos, vacias, ok = [], 0, 0
    for fname, registro in registros.items():
        if not registro:
            continue
        wbv = openpyxl.load_workbook(os.path.join(carpeta, fname),
                                     data_only=True)
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append(f'{fname}:{hoja}!{coord}: hoja ausente')
                continue
            v = wbv[hoja][coord].value
            if v is None:
                if '""' in formula:
                    vacias += 1
                else:
                    fallos.append(f'{fname}:{hoja}!{coord}: sin cache '
                                  f'({formula[:70]})')
            else:
                ok += 1
    return {'con_valor': ok, 'vacias_por_diseno': vacias, 'fallos': fallos}


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return f'ERR:{type(e).__name__}: {e}'


def _copia(carpeta, fname, demos, etiqueta):
    os.makedirs(demos, exist_ok=True)
    dst = os.path.join(demos, f'{etiqueta}-{fname}')
    shutil.copy2(os.path.join(carpeta, fname), dst)
    return dst


def demo_liquidacion(carpeta, demos):
    """§6 (m6) — el arqueo no se puede demostrar en un kit que no tiene cajón.

    En el modelo POR EVENTOS la cuenta que sostiene el fichero del dinero es
    PENDIENTE = TOTAL FACTURA − anticipo − cobrado, y el ESTADO que sale de
    ella. Se demuestran los dos casos que importan y que un cliente distingue
    de un vistazo:

      · «liquidacion-cuadra»  — se cobra todo el saldo → PENDIENTE 0 y
        ESTADO «Cobrado».
      · «liquidacion-vencida» — no se cobra nada y el vencimiento ya pasó →
        PENDIENTE = saldo y ESTADO «VENCIDO», que es el aviso de reclamar.

    Las celdas se localizan por su RÓTULO (las constantes `ETIQ_EV_*` de
    motor.py), nunca por coordenada: el constructor puede mover una fila y la
    demostración tiene que seguir midiendo lo mismo.
    """
    from pycel import ExcelCompiler
    fname = motor.CTX.get('f_caja')
    if not fname:
        return []
    ruta = os.path.join(carpeta, fname)
    wb0 = openpyxl.load_workbook(ruta)
    hoja = None
    for ws in wb0.worksheets:
        if motor.fila_liquidacion(ws):
            hoja = ws.title
            break
    if hoja is None:
        return [{'caso': 'liquidacion-cuadra', 'ref': fname,
                 'entradas': 'ninguna', 'esperado': 'hoja de liquidación',
                 'obtenido': 'el fichero de COBROS no tiene ninguna hoja con '
                             '«TOTAL FACTURA» + «PENDIENTE DE COBRO»',
                 'ok': False}]
    ws0 = wb0[hoja]

    def fila(etq):
        return (motor._buscar(ws0, etq, col=1)
                or motor._buscar(ws0, etq, col=2))

    filas = {k: fila(v) for k, v in (
        ('presupuesto', motor.ETIQ_EV_PRESUPUESTO),
        ('extras', motor.ETIQ_EV_EXTRAS),
        ('base10', motor.ETIQ_EV_BASE10), ('base21', motor.ETIQ_EV_BASE21),
        ('total', motor.ETIQ_EV_TOTAL), ('anticipo', motor.ETIQ_EV_ANTICIPO),
        ('saldo', motor.ETIQ_EV_SALDO), ('cobrado', motor.ETIQ_EV_COBRADO),
        ('pendiente', motor.ETIQ_EV_PENDIENTE),
        ('vencimiento', motor.ETIQ_EV_VENCIMIENTO),
        ('estado', motor.ETIQ_EV_ESTADO))}
    faltan = sorted(k for k, v in filas.items() if not v)
    if faltan:
        return [{'caso': 'liquidacion-cuadra', 'ref': f'{fname}:{hoja}',
                 'entradas': 'ninguna', 'esperado': 'los 11 rótulos del molde',
                 'obtenido': 'faltan rótulos: ' + ', '.join(faltan),
                 'ok': False}]

    hoy = datetime.date.today()
    # base 10 % = 8.000 y base 21 % = 2.000 → 8.800 + 2.420 = 11.220 €.
    # Anticipo del 40 % sobre la base (4.000) → saldo 7.220 €.
    casos = []
    for etiqueta, cobrado, venc, esp_pend, esp_estado in (
            ('liquidacion-cuadra', 7220,
             hoy + datetime.timedelta(days=30), 0, motor.EV_COBRADO),
            ('liquidacion-vencida', 0,
             hoy - datetime.timedelta(days=10), 7220, motor.EV_VENCIDO)):
        dst = _copia(carpeta, fname, demos, etiqueta)
        wb = openpyxl.load_workbook(dst)
        ws = wb[hoja]
        for clave, valor in (('presupuesto', 10000), ('extras', 0),
                             ('base10', 8000), ('base21', 2000),
                             ('anticipo', 4000), ('cobrado', cobrado)):
            ws.cell(row=filas[clave], column=3).value = valor
        ws.cell(row=filas['vencimiento'], column=3).value = venc
        wb.save(dst)
        xl = ExcelCompiler(filename=dst)
        pend = _ev(xl, "'{}'!C{}".format(hoja, filas['pendiente']))
        estado = _ev(xl, "'{}'!C{}".format(hoja, filas['estado']))
        casos.append({
            'caso': etiqueta,
            'ref': '{}:{}:C{}/C{}'.format(fname, hoja, filas['pendiente'],
                                          filas['estado']),
            'entradas': {'presupuesto (base)': 10000, 'extras': 0,
                         'base 10 %': 8000, 'base 21 %': 2000,
                         'anticipo cobrado': 4000,
                         'cobrado tras el evento': cobrado,
                         'vencimiento': venc.isoformat(),
                         'hoy': hoy.isoformat()},
            'total_factura': _ev(xl, "'{}'!C{}".format(hoja, filas['total'])),
            'saldo_tras_anticipo': _ev(xl, "'{}'!C{}".format(hoja,
                                                            filas['saldo'])),
            'esperado': '{} / {}'.format(esp_pend, esp_estado),
            'obtenido': '{} / {}'.format(pend, estado),
            'ok': (isinstance(pend, (int, float))
                   and abs(pend - esp_pend) <= motor.EV_TOLERANCIA
                   and estado == esp_estado),
            'copia': dst})
    return casos


def demo_arqueo(carpeta, demos):
    """§6 — fondo 150, efectivo contado 1.150, tarjetas 800, Z 1.800 → 0."""
    from pycel import ExcelCompiler
    fname = motor.CTX.get('f_caja')
    if not fname:
        return []
    # m6 — en el modelo POR EVENTOS no hay fondo, ni cajón, ni Z del TPV: la
    # demostración exigible es otra. Se DELEGA en vez de devolver [] para que
    # `fallos` siga cazando «§6: no se pudo ejecutar ninguna demostración».
    if motor.es_modelo_eventos():
        return demo_liquidacion(carpeta, demos)
    casos = []
    for etiqueta, z, esperado in (('arqueo-cuadra', 1800, 0),
                                  ('arqueo-descuadra', 1790, 10)):
        dst = _copia(carpeta, fname, demos, etiqueta)
        wb = openpyxl.load_workbook(dst)
        ap, ci = wb['Apertura de Caja'], wb['Cierre de Caja']
        r_fondo = motor._buscar(ap, motor.ETIQ_FONDO, col=2)
        ap.cell(row=r_fondo, column=3).value = 150
        for etq, cant in (('500 €', 2), ('100 €', 1), ('50 €', 1)):
            r = motor._buscar(ci, etq, col=1)
            ci.cell(row=r, column=2).value = cant
        r_tar = motor._buscar(ci, 'Total Tarjetas (Visa/MC)', col=1)
        r_z = motor._buscar(ci, motor.ETIQ_Z, col=1)
        r_desc = motor._buscar(ci, 'DESCUADRE', col=1)
        r_tot = motor._buscar(ci, 'TOTAL FACTURADO', col=1)
        r_ef = motor._buscar(ci, motor.ETIQ_EFECTIVO, col=1)
        ci.cell(row=r_tar, column=3).value = 800
        ci.cell(row=r_z, column=3).value = z
        wb.save(dst)
        xl = ExcelCompiler(filename=dst)
        obtenido = _ev(xl, f"'Cierre de Caja'!C{r_desc}")
        casos.append({
            'caso': etiqueta,
            'ref': f"{fname}:Cierre de Caja:C{r_desc}",
            'entradas': {f'Apertura de Caja!C{r_fondo}': 150,
                         'Cierre de Caja!B(500/100/50 €)': '2/1/1 = 1.150 €',
                         f'Cierre de Caja!C{r_tar}': 800,
                         f'Cierre de Caja!C{r_z}': z},
            'efectivo_contado': _ev(xl, f"'Cierre de Caja'!C{r_ef}"),
            'total_facturado': _ev(xl, f"'Cierre de Caja'!C{r_tot}"),
            'esperado': esperado, 'obtenido': obtenido,
            'ok': obtenido == esperado, 'copia': dst})
    return casos


def demo_contador(carpeta, demos):
    """§6 — 25 tareas, 3 N/A, 22 ✓ → «22 de 22»."""
    from pycel import ExcelCompiler
    mejor = None
    for fname in motor.CTX.get('con_checklist', ()):
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wb.worksheets:
            g = motor.geometria(ws)
            if not g or not g['contador']:
                continue
            tareas = [r for r in range(g['hr'] + 1, g['contador'])
                      if isinstance(ws.cell(row=r, column=2).value, str)
                      and ws.cell(row=r, column=2).value.strip()
                      and not motor.es_fila_seccion(ws, r)]
            if len(tareas) == 25:
                mejor = (fname, ws.title, g, tareas)
                break
            if not mejor or len(tareas) > len(mejor[3]):
                mejor = (fname, ws.title, g, tareas)
        if mejor and len(mejor[3]) == 25:
            break
    if not mejor:
        return []
    fname, hoja, g, tareas = mejor
    dst = _copia(carpeta, fname, demos, 'contador')
    wb = openpyxl.load_workbook(dst)
    ws = wb[hoja]
    for i, r in enumerate(tareas):
        ws.cell(row=r, column=g['marca']).value = 'N/A' if i < 3 else '✓'
    wb.save(dst)
    xl = ExcelCompiler(filename=dst)
    col_num = motor.get_column_letter(g['marca'] - 2)
    col_den = motor.get_column_letter(g['marca'])
    num = _ev(xl, f"'{hoja}'!{col_num}{g['contador']}")
    den = _ev(xl, f"'{hoja}'!{col_den}{g['contador']}")
    esperado = (len(tareas) - 3, len(tareas) - 3)
    return [{'caso': 'contador-honesto',
             'ref': f'{fname}:{hoja}:{col_num}{g["contador"]}/'
                    f'{col_den}{g["contador"]}',
             'entradas': f'{len(tareas)} tareas · 3 N/A · {len(tareas) - 3} ✓',
             'esperado': f'{esperado[0]} de {esperado[1]}',
             'obtenido': f'{num} de {den}',
             'ok': (num, den) == esperado, 'copia': dst}]


def demo_07(carpeta, demos):
    """§6 — 8 tareas escritas en la plantilla en blanco → «de 8»."""
    from pycel import ExcelCompiler
    for fname in sorted(motor.CTX.get('con_checklist', ())):
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        # La geometría ▸ se exige aquí y no sólo el título: los sinónimos de
        # §2.3 («Por Zona») también los llevan hojas del molde P4 (catering,
        # chocolatería, heladería) y sin este filtro la demostración se
        # llevaría una de ellas y reventaría en `g['hr']` con `g = None`.
        hojas = [ws.title for ws in wb.worksheets
                 if ws.title in motor.PLANTILLA_07 and motor.geometria(ws)]
        if len(hojas) < 3:
            continue
        dst = _copia(carpeta, fname, demos, 'plantilla-8-tareas')
        wb = openpyxl.load_workbook(dst)
        ws = wb[hojas[0]]
        g = motor.geometria(ws)
        libres = [r for r in range(g['hr'] + 1, g['contador'])
                  if not motor.es_fila_seccion(ws, r)]
        for r in libres:
            ws.cell(row=r, column=2).value = None
        for i, r in enumerate(libres[:8]):
            ws.cell(row=r, column=2).value = f'Tarea propia {i + 1}'
        wb.save(dst)
        xl = ExcelCompiler(filename=dst)
        col_den = motor.get_column_letter(g['marca'])
        den = _ev(xl, f"'{hojas[0]}'!{col_den}{g['contador']}")
        return [{'caso': 'denominador-07',
                 'ref': f'{fname}:{hojas[0]}:{col_den}{g["contador"]}',
                 'entradas': '8 tareas escritas, ninguna marcada',
                 'esperado': 8, 'obtenido': den, 'ok': den == 8,
                 'copia': dst}]
    return []


def demo_p4(carpeta, nombres, demos):
    """(l) — molde P4: hoja RECIÉN IMPRESA → «0 de N», N = tareas reales.

    El contador del molde P4 contaba los rótulos: la cabecera «Nº | Tarea | …»
    se repite en cada sección, así que `COUNTIF(B,"?*")` sumaba esas filas al
    denominador y `COUNTIF(E,"✓")` sumaba sus «✓» al numerador. Medido en
    heladería, 01-apertura-cierre.xlsx:'Apertura'!C29/E29: la hoja sin marcar
    nada anunciaba «2 de 19» con 17 tareas. La demostración NO toca el fichero
    —la gracia es que sea el que se entrega— y compara contra el recuento
    hecho a mano, publicando además lo que habría dicho la fórmula vieja.
    """
    from pycel import ExcelCompiler
    for fname in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
        for ws in wb.worksheets:
            g = motor.geometria_p4(ws)
            if not g or not g['contador']:
                continue
            lo, hi = g['hr'] + 1, g['contador'] - 1
            if hi < lo:
                continue
            reales = rotulos = marcadas = ingenuo_num = 0
            for r in range(lo, hi + 1):
                v = ws.cell(row=r, column=2).value
                marca = ws.cell(row=r, column=g['marca']).value
                if marca == motor.MARCA_OK:
                    ingenuo_num += 1
                if not isinstance(v, str) or not v.strip():
                    continue
                if v.strip() == 'Tarea':
                    rotulos += 1
                    continue
                reales += 1
                if marca == motor.MARCA_OK:
                    marcadas += 1
            if not reales or not rotulos:
                continue        # sin cabeceras repetidas no demuestra nada
            dst = _copia(carpeta, fname, demos, 'p4-recien-impresa')
            xl = ExcelCompiler(filename=dst)
            cn = motor.get_column_letter(g['marca'] - 2)
            cd = motor.get_column_letter(g['marca'])
            num = _ev(xl, "'{}'!{}{}".format(ws.title, cn, g['contador']))
            den = _ev(xl, "'{}'!{}{}".format(ws.title, cd, g['contador']))
            return [{
                'caso': 'contador-p4-sin-rotulos',
                'ref': '{}:{}:{}{}/{}{}'.format(fname, ws.title, cn,
                                                g['contador'], cd,
                                                g['contador']),
                'entradas': ('hoja tal y como se entrega · {} tareas reales, '
                             '{} filas de rótulo repetido, {} marcadas ✓'
                             .format(reales, rotulos, marcadas)),
                'formula_vieja_habria_dicho': '{} de {}'.format(
                    ingenuo_num, reales + rotulos),
                'esperado': '{} de {}'.format(marcadas, reales),
                'obtenido': '{} de {}'.format(num, den),
                'ok': (num, den) == (marcadas, reales), 'copia': dst}]
    return []


def gate_recuento(carpeta, nombres):
    """TEC-R2-01/COM-R2-01 — cuántas tareas entrega el kit, DE VERDAD.

    La landing publicaba «121 tareas en el kit completo (recuento v2.0)» y el
    recuento no se había hecho: 121 era el fichero 01 de la v1.1. La cifra sale
    de los mismos denominadores que enseña el Excel, así que la página de venta
    y el producto no pueden volver a divergir sin que este gate lo cante.

    T-03 (tanda 4) — se suman TODAS las hojas de checklist del producto: el
    molde ▸, el molde P4 y las plantillas con denominador por fórmula. Antes
    sólo sumaba el molde ▸ y en los cinco kits P4 (catering, chocolatería,
    heladería, hotel, restaurante-creativo) devolvía 56 —las tareas de 08 y 09—
    ignorando los cientos de los 01-07: hotel habría anunciado 56 tareas en vez
    de 636. Los rótulos repetidos de cada sección del molde P4 NO cuentan: los
    descuenta la propia fórmula del denominador. `recuento_por_molde` separa de
    dónde sale cada cifra para que el número de la landing sea auditable.
    """
    por_fichero, total = {}, 0
    por_molde = {'checklist_▸': 0, 'p4': 0}
    por_hoja = []
    for n in nombres:
        wbv = openpyxl.load_workbook(os.path.join(carpeta, n), data_only=True)
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        n_tareas = 0
        for ws in wb.worksheets:
            molde = 'checklist_▸'
            g = motor.geometria(ws)
            if not g:
                g, molde = motor.geometria_p4(ws), 'p4'
            if not g or not g.get('contador'):
                continue
            v = wbv[ws.title].cell(row=g['contador'], column=g['marca']).value
            try:
                v = int(v or 0)
            except (TypeError, ValueError):
                # Sin cache el denominador vendría como fórmula: se anota en
                # vez de contarlo como 0 en silencio.
                por_hoja.append({'fichero': n, 'hoja': ws.title,
                                 'molde': molde, 'tareas': None,
                                 'celda': '{}{}'.format(
                                     motor.get_column_letter(g['marca']),
                                     g['contador']),
                                 'aviso': 'denominador sin valor cacheado'})
                continue
            n_tareas += v
            por_molde[molde] += v
            por_hoja.append({'fichero': n, 'hoja': ws.title, 'molde': molde,
                             'celda': '{}{}'.format(
                                 motor.get_column_letter(g['marca']),
                                 g['contador']),
                             'tareas': v})
        por_fichero[n] = n_tareas
        total += n_tareas
    return {'total': total, 'por_fichero': por_fichero,
            'recuento_por_molde': por_molde, 'por_hoja': por_hoja}


#: T-02 — censo de las tareas que ENCIENDEN el TPV en todo el producto. La del
#: fichero de negocio es el hito legítimo; la del de caja la reescribe el motor;
#: cualquier otra (heladería 01!B24, pastelería 01!B29) es OTRA CAPA y se deja,
#: pero queda registrada aquí para que nadie tenga que volver a buscarla.
RX_TPV_GATE = re.compile(
    r'(?i)^\s*encender\b.*?(tpv|\bpos\b|caja registradora)')


def gate_tpv(carpeta, nombres):
    encendidos = []
    for n in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        for ws in wb.worksheets:
            g = motor.geometria(ws) or motor.geometria_p4(ws)
            if not g:
                continue
            for r in range(g['hr'] + 1, (g.get('contador') or ws.max_row)):
                v = ws.cell(row=r, column=2).value
                if not isinstance(v, str) or not RX_TPV_GATE.match(v):
                    continue
                if n == motor.CTX.get('f_negocio'):
                    papel = 'negocio (hito legítimo)'
                elif n == motor.CTX.get('f_caja'):
                    # m6 — T-02 («el TPV se enciende UNA vez, en el fichero de
                    # negocio») es una regla del modelo de MOSTRADOR: nace de
                    # que el 08 y el 09 encendían el mismo TPV con dos horas
                    # distintas. En el modelo POR EVENTOS el fichero del dinero
                    # no abre ningún turno de TPV, así que el gate no le exige
                    # nada; si alguien escribe ahí una tarea de encender un TPV
                    # (la barra opcional en efectivo), se anota y se deja.
                    papel = ('cobros (modelo por eventos: T-02 no aplica)'
                             if motor.es_modelo_eventos()
                             else 'caja (T-02 NO aplicado: revisar)')
                elif n == motor.CTX.get('f_areas'):
                    papel = 'areas (otra capa: se deja)'
                else:
                    papel = 'otro fichero (otra capa: se deja)'
                encendidos.append({'ref': '{}:{}!B{}'.format(n, ws.title, r),
                                   'papel': papel, 'texto': v})
    return {'tareas_encender_tpv': encendidos,
            'fuera_del_fichero_de_negocio': [
                d['ref'] for d in encendidos
                if not d['papel'].startswith('negocio')]}


def gate_metadata(carpeta, nombres):
    """m5 — `subject`, `title`, `creator` y `keywords` en TODOS los ficheros.

    Sin este gate, m5 no sería auditable: el defecto que corrige (26 ficheros
    del molde P4 y de los BONUS publicándose con `subject` «… · v1.1» dentro de
    un producto v2.0) vivió cuatro tandas precisamente porque ningún gate miraba
    las propiedades de los ficheros fuera del molde ▸ — el censo mira las hojas.
    """
    esperado_sub = '{} · v2.0'.format(
        motor.CTX.get('sufijo') or 'Kit de Tareas Recurrentes Pro')
    esperado_kw = motor.keywords_del_kit()
    mal, detalle, propias = [], [], []
    for n in nombres:
        p = openpyxl.load_workbook(os.path.join(carpeta, n)).properties
        detalle.append({'fichero': n, 'title': p.title, 'subject': p.subject,
                        'creator': p.creator, 'keywords': p.keywords})
        if p.subject != esperado_sub:
            mal.append(f'{n}: subject = {p.subject!r} (esperado '
                       f'{esperado_sub!r})')
        if p.creator != 'AI Chef Pro':
            mal.append(f'{n}: creator = {p.creator!r}')
        if not motor.keywords_ok(p.keywords):
            mal.append(f'{n}: keywords = {p.keywords!r} — no sigue la '
                       f'convención «…, {motor.COLA_KEYWORDS}»')
        elif p.keywords != esperado_kw:
            # AVISO, no fallo: `kit-tareas-pasteleria` lleva keywords escritas
            # a mano, más ricas que las derivadas. Se respetan y se anotan.
            propias.append(f'{n}: keywords propias {p.keywords!r} '
                           f'(el derivado sería {esperado_kw!r}) — se respetan')
        if not p.title or not p.title.endswith(
                ' · ' + (motor.CTX.get('sufijo') or '')):
            mal.append(f'{n}: title = {p.title!r} — no acaba en « · '
                       f'{motor.CTX.get("sufijo")}»')
    return {'subject_esperado': esperado_sub, 'keywords_esperado': esperado_kw,
            'incoherentes': mal, 'keywords_propias': propias,
            'por_fichero': detalle}


def gate_dv_y_bio(carpeta, nombres):
    """DV «✓,—,N/A» en TODA hoja de checklist y bio en TODA Instrucciones.

    TEC-R2-08: también se censan las hojas FUERA de alcance. Aplicar sólo el
    motor a los 08/09 de un hermano (el kit de hotel) deja el producto con dos
    listas de desplegable conviviendo —4 hojas con «✓,—,N/A» y 48 con «✓,✗,—»—
    y el gate lo daba por verde porque sólo miraba lo que había tocado.
    """
    dv_mal, sin_bio, hojas, con_bio = [], [], 0, 0
    listas_kit, hojas_p4 = {}, 0
    sin_instrucciones, version_mal = [], []
    for n in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        for ws in wb.worksheets:
            for dv in ws.data_validations.dataValidation:
                if dv.type == 'list':
                    listas_kit.setdefault(dv.formula1, []).append(
                        f'{n}:{ws.title}')
        # R3-e — las hojas del molde P4 también se auditan: son las que dejaban
        # el producto con dos desplegables y dos semánticas de conteo.
        for ws in wb.worksheets:
            if not motor.geometria_p4(ws):
                continue
            hojas_p4 += 1
            listas = {dv.formula1 for dv in ws.data_validations.dataValidation
                      if dv.type == 'list'}
            if listas and listas != {motor.DV_LISTA}:
                dv_mal.append(f'{n}:{ws.title} (molde P4): DV = '
                              f'{sorted(listas)} (esperada {motor.DV_LISTA})')
        # (j) — la autoría y la versión se censan en TODOS los ficheros del
        # producto, también los que quedan FUERA de alcance. Antes el censo iba
        # detrás del `continue` de abajo y los dos calendarios que no son ni
        # checklist ▸ ni molde P4 —los BONUS-02 de catering y de hotel— se
        # publicaban dentro de un producto v2.0 diciendo «Versión 1.1 · agosto
        # 2026» y sin la línea de autoría, sin que ningún gate lo mirase.
        ins = wb['Instrucciones'] if 'Instrucciones' in wb.sheetnames else None
        if ins is None:
            # No se inventa una hoja: se anota y decide el orquestador.
            sin_instrucciones.append(n)
        else:
            texto = ' '.join(str(c.value) for row in ins.iter_rows()
                             for c in row if c.value)
            if motor.RX_BIO.search(texto):
                con_bio += 1
            else:
                sin_bio.append(n)
            if motor.version_line() not in texto:
                version_mal.append(f'{n}:Instrucciones: no dice '
                                   f'«{motor.version_line()}»')
        if not motor.en_alcance(wb):
            continue
        for ws in wb.worksheets:
            if not motor.geometria(ws):
                continue
            hojas += 1
            listas = {dv.formula1 for dv in ws.data_validations.dataValidation
                      if dv.type == 'list'}
            if listas != {motor.DV_LISTA}:
                dv_mal.append(f'{n}:{ws.title}: DV = {sorted(listas)} '
                              f'(esperada {motor.DV_LISTA})')
    mezcla = []
    if len(listas_kit) > 1:
        mezcla = [f'{f} en {len(h)} hojas (p.ej. {h[0]})'
                  for f, h in sorted(listas_kit.items())]
    return {'hojas_checklist': hojas, 'hojas_p4': hojas_p4,
            'dv_incorrectas': dv_mal,
            'instrucciones_con_bio': con_bio, 'sin_bio': sin_bio,
            'sin_hoja_instrucciones': sin_instrucciones,
            'version_desfasada': version_mal,
            'listas_dv_del_producto': sorted(listas_kit),
            'aviso_dv_mezcladas': mezcla}


def gate_ortografia(carpeta, nombres):
    """CB-E1 §1.3 — 0 lemas de `LEX_TILDES` vivos en celdas y en pestañas.

    Baseline a batir (medido sobre `dl/` el 2026-08-29, los 7 kits de la
    sub-familia): 845 palabras · 653 celdas · 6 pestañas.

    El gate NO es simétrico con el paso: además de contar lo que queda mal,
    EMITE la lista de palabras AMBIGUAS («esta/está», «solo/sólo», «mas/más»,
    «como/cómo»…), que el diccionario no toca a propósito y que resuelve el
    módulo de contenido del kit celda a celda (§1.2 CB-E1). Sin esa lista, «0
    lemas» se leería como «ortografía perfecta», que es otra cosa.
    """
    vivos, ambiguas = [], []
    for n in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        for hoja, celda, texto in motor.palabras_sin_tilde(wb):
            vivos.append({'ref': f'{n}!{hoja}!{celda}', 'texto': texto[:160]})
        for a in motor.ambiguas_del_libro(wb):
            ambiguas.append({'ref': f"{n}!{a['hoja']}!{a['celda']}",
                             'palabra': a['palabra'], 'texto': a['texto']})
    return {'aplica': motor.sub_cb(),
            'lemas_vivos': vivos,
            'ambiguas': ambiguas,
            'resumen': f'{len(vivos)} celdas con lema de LEX_TILDES · '
                       f'{len(ambiguas)} ocurrencias ambiguas para el módulo '
                       'de contenido'}


def gate_contadores(carpeta, nombres):
    """CB-E2 §1.3 — toda hoja con geometría de checklist tiene contador.

    `kit-tareas-sushi-bar/09-plantilla-personalizable.xlsx:'Plantilla en
    Blanco'` entraba en alcance con `contador = None` y 0 fórmulas mientras sus
    «Instrucciones» estrenaban los bloques que lo explican.
    """
    sin, con = [], 0
    for n in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        for ws in wb.worksheets:
            g = motor.geometria(ws)
            if not g:
                continue
            if g['contador'] is None:
                sin.append(f'{n}!{ws.title}: geometría de checklist SIN fila '
                           'de totales')
            else:
                con += 1
    return {'con_contador': con, 'sin_contador': sin}


def gate_moldes(carpeta, nombres):
    """§1.3 `molde` — cada fichero, clasificado; y cada hoja, con su tipo."""
    por_fichero, sin_clasificar = {}, []
    for n in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        recon = motor.hojas_reconocidas(wb)
        p4 = [ws.title for ws in wb.worksheets
              if ws.title not in recon and motor.geometria_p4(ws)]
        sueltas = [ws.title for ws in wb.worksheets
                   if ws.title != 'Instrucciones' and ws.title not in recon
                   and ws.title not in p4]
        por_fichero[n] = {'reconocidas': recon, 'p4': p4, 'sin_molde': sueltas}
        sin_clasificar += [f'{n}!{t}' for t in sueltas]
    return {'por_fichero': por_fichero, 'hojas_sin_molde': sin_clasificar}


def gate_precargado(carpeta):
    """§6 — «08: ninguna tarea sin Responsable/Hora» (DOM-06/TEC-06).

    R3-b — el gate AUDITA QUE ESTÁ AUDITANDO EL FICHERO CORRECTO. En bar y en
    dark-kitchen daba verde («48 tareas, 0 huecos») sobre 01-apertura-cierre,
    que ya venía con sus responsables puestos, mientras el 08 real salía con
    D5:E22 vacías en las 18 tareas. Un gate que no comprueba su propio sujeto
    no es un gate: por él pasó dark-kitchen con exit 0 estando roto.
    """
    fname = motor.CTX.get('f_negocio')
    if not fname:
        # CB-E7 — hay kits que no tienen fichero de negocio NI de caja: no es
        # que la detección falle, es que el producto no trae 08/09 (sushi-bar,
        # asador y chef-privado; medido sobre `dl/` el 2026-08-29). Con la
        # bandera puesta el gate pasa de FALLO a INFORMATIVO: seguir sacando
        # `exit 1` por «DOM-06 no se ha aplicado a nada» convertía el informe
        # de esos kits en un rojo permanente que tapa los fallos de verdad.
        # La bandera exige además que ningún fichero traiga firma del dinero,
        # así que un fallo REAL de detección sigue saliendo en rojo por aquí.
        if motor.CTX.get('sin_caja'):
            return {'fichero': None, 'tareas': 0, 'huecos': [],
                    'firma_ok': True, 'huecos_firma': [], 'sin_caja': True,
                    'informativo': [
                        'CB-E7 — el kit no tiene fichero de NEGOCIO ni de '
                        'CAJA (ni firma de recuento, registro mensual, '
                        'liquidación o registro de eventos en ninguno de sus '
                        'ficheros): DOM-06 no aplica a este producto.']}
        return {'fichero': None, 'tareas': 0, 'huecos': [],
                'firma_ok': False, 'sin_caja': False,
                'huecos_firma': ['no se ha identificado el fichero de NEGOCIO '
                                 'del kit: DOM-06 no se ha aplicado a nada']}
    wb = openpyxl.load_workbook(os.path.join(carpeta, fname))
    papel, detalle = motor.papel_del_fichero(wb)
    firma = []
    if papel != 'negocio':
        firma.append(f'{fname}: el gate está auditando un fichero que NO es el '
                     f'de negocio (papel detectado: {papel}; {detalle})')
    if len(detalle.get('con_notas', [])) != 2:
        firma.append(f'{fname}: se esperaban 2 checklists de apertura/cierre '
                     f'con columna «Notas» y hay '
                     f'{len(detalle.get("con_notas", []))} '
                     f'({detalle.get("con_notas")})')
    huecos, tareas = [], 0
    for ws in wb.worksheets:
        g = motor.geometria(ws)
        if not g:
            continue
        col_r = g['cols'].get('Responsable')
        col_h = motor._col_tiempo(g['cols'])
        for r in range(g['hr'] + 1, g['ultima'] + 1):
            if not isinstance(ws.cell(row=r, column=1).value, int):
                continue
            tareas += 1
            for nombre, c in (('Responsable', col_r), ('Hora', col_h)):
                if c and not ws.cell(row=r, column=c).value:
                    huecos.append(f'{fname}:{ws.title}!'
                                  f'{motor.get_column_letter(c)}{r}: '
                                  f'{nombre} vacío')
    return {'fichero': fname, 'tareas': tareas, 'huecos': huecos,
            'firma_ok': not firma, 'huecos_firma': firma,
            'checklists_con_notas': detalle.get('con_notas')}


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-1200:])
    salida = (r.stdout.strip() or r.stderr.strip()).splitlines()
    return {'exit': r.returncode, 'salida': salida[-8:]}


def inventario(carpeta, nombres):
    fuera = []
    for n in nombres:
        wb = openpyxl.load_workbook(os.path.join(carpeta, n))
        hojas = []
        for ws in wb.worksheets:
            hojas.append({
                'hoja': ws.title, 'dim': ws.dimensions,
                'dv': [f'{dv.type}:{dv.formula1}'
                       for dv in ws.data_validations.dataValidation],
                'cf': len(list(ws.conditional_formatting)),
                'print_area': ws.print_area,
                'orient': ws.page_setup.orientation,
                'a4': ws.page_setup.paperSize == 9,
                'protegida': bool(ws.protection.sheet),
                'formulas': sum(1 for row in ws.iter_rows() for c in row
                                if c.data_type == 'f'),
            })
        fuera.append({'fichero': n, 'title': wb.properties.title,
                      'subject': wb.properties.subject, 'hojas': hojas})
    return fuera


# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 de la familia Kit de Tareas')
    ap.add_argument('--producto', default='kit-tareas')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--solo', default='motor,contenido',
                    help='etapas: motor, contenido')
    ap.add_argument('--json', default=None)
    # m7 — carpeta de origen alternativa. Sirve para probar el motor contra un
    # producto que TODAVÍA no está en `dl/` (el 09 nuevo de catering se
    # construye fuera y sustituye al 09 viejo) sin tener que tocar `dl/**`, que
    # la SPEC declara intocable. Sólo con --dry-run: con --origen y sin
    # --dry-run el script escribiría en la carpeta que le pasen, que es
    # exactamente lo que la regla dura impide.
    ap.add_argument('--origen', default=None,
                    help='carpeta de origen alternativa (exige --dry-run)')
    ap.add_argument('--sin-idempotencia', action='store_true')
    ap.add_argument('--sin-demos', action='store_true')
    args = ap.parse_args()

    pid = args.producto
    if args.origen and not args.dry_run:
        raise SystemExit('ABORTADO: --origen sólo se admite con --dry-run.')
    origen = args.origen or os.path.join(DL, pid)
    destino = os.path.join(SCRATCH, 'dryrun-kt', pid)
    idem_dir = os.path.join(SCRATCH, 'dryrun-kt', pid + '-idem')
    demos_dir = os.path.join(SCRATCH, 'dryrun-kt', '_demos', pid)

    if not args.dry_run and os.environ.get('KIT_TAREAS_APPLY') != '1':
        raise SystemExit(
            f'ABORTADO: sin --dry-run este script escribiría en {origen}, que '
            'la SPEC declara intocable hasta que la ronda 2 dé verde. Usa '
            '--dry-run (o KIT_TAREAS_APPLY=1 si eres el orquestador y ya '
            'tienes el visto bueno).')

    respaldo = None
    if args.dry_run:
        preparar_copia(origen, destino)
        carpeta = destino
    else:
        carpeta = origen
        respaldo = os.path.join(
            SCRATCH, f'{pid}.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(origen, respaldo)
        log(f'  respaldo previo de los entregables: {respaldo}')

    etapas = {e.strip() for e in args.solo.split(',') if e.strip()}
    contenido = None
    modulo = 'contenido_' + pid.replace('-', '_')
    if 'contenido' in etapas:
        try:
            contenido = importlib.import_module(modulo)
            log(f'  {modulo} cargado')
        except ImportError:
            log(f'  {modulo}.py no existe — sólo motor')

    # m6/m8 — el 09 de catering se reconstruye ANTES de censar la carpeta y
    # antes del contexto: el modelo del dinero ('mostrador' o 'eventos') se
    # decide leyendo ese fichero.
    sust09 = sustituir_09_catering(carpeta, pid)
    if sust09:
        log('  09 de catering: construido {}{}'.format(
            sust09['construido'],
            ' y borrado ' + sust09['borrado'] if sust09['borrado'] else ''))

    nombres = ficheros_de(carpeta)
    log(f'\n== 1/7 · contexto del kit ({len(nombres)} xlsx) ==')
    try:
        ctx = motor.contexto(
            carpeta, nombres,
            lambda f: openpyxl.load_workbook(os.path.join(carpeta, f)),
            producto=pid)
    except (motor.KitAmbiguo, motor.MoldeDesconocido) as e:
        # R3-a — el motor NO adivina el papel de un fichero. Se aborta con el
        # informe escrito, para que el orquestador vea por qué.
        # CB-E6 — y tampoco guarda un producto cuyo molde no reconoce. El
        # aborto ocurre ANTES del bucle de proceso, así que no se ha escrito ni
        # un fichero: en dry-run la copia de trabajo queda idéntica al origen.
        log('\nABORTADO — ' + str(e))
        if args.json:
            os.makedirs(os.path.dirname(os.path.abspath(args.json)),
                        exist_ok=True)
            with open(args.json, 'w', encoding='utf-8') as fh:
                json.dump({'producto': pid, 'version': '2.0', 'spec': SPEC,
                           'modo': 'dry-run' if args.dry_run else 'produccion',
                           'motivo': type(e).__name__,
                           'abortado': str(e), 'fallos': [str(e)], 'exit': 2},
                          fh, ensure_ascii=False, indent=1)
        return 2
    ctx['producto'] = pid
    motor.CTX['producto'] = pid
    log(f"  kit «{ctx['kit']}» · en alcance {len(ctx['ficheros'])}/"
        f"{len(nombres)} · negocio={ctx['f_negocio']} caja={ctx['f_caja']} "
        f"areas={ctx['f_areas']} · hora ancla {ctx['hora_apertura']} · "
        f"cierre «{ctx['literal_cierre']}»")

    log(f'\n== 2/7 · post-proceso ==')
    motor.SOLAPES.clear()
    informe_ficheros, registros = [], {}
    for fname in nombres:
        registros[fname] = procesar(carpeta, fname, etapas, contenido,
                                    informe_ficheros)
        f = informe_ficheros[-1]
        if f['en_alcance']:
            estado_txt = (f"{len(f['cambios'])} cambios, "
                          f"{f['formulas_nuevas']} fórmulas")
        elif f['guardado']:
            # R3-e — «fuera de alcance» ya no significa «intacto»: el molde P4
            # recibe DV, contador, CF y bio sin entrar en el molde ▸.
            estado_txt = (f"fuera del molde ▸ · molde P4: "
                          f"{len(f['cambios'])} cambios, "
                          f"{f['formulas_nuevas']} fórmulas")
        else:
            estado_txt = 'FUERA DE ALCANCE (intacto)'
        log(f'  {fname}: {estado_txt}')

    # R3-f — se fotografía ANTES de la 2.ª pasada, que volvería a acumular.
    # T-05 — y se DEDUPLICA: cuando el módulo de contenido cambia la estructura
    # del libro, `procesar` vuelve a pasar el motor sobre el mismo fichero y
    # cada banda se medía dos veces (pizzería). Sólo afectaba al JSON del
    # informe, pero falseaba cualquier censo agregado. Si una misma banda
    # apareciera con DOS medidas distintas se conservan las dos: eso no sería
    # ruido, sería el motor dando resultados inestables.
    #
    # m2 (tanda 5) — la clave de T-05 incluía la REFERENCIA DE CELDA
    # («Cierre Cocina!A28»), y el módulo de contenido que inserta una fila por
    # encima de la banda mueve esa referencia: la 2.ª pasada del motor mide la
    # MISMA banda en A29 y las dos entradas sobrevivían a la deduplicación.
    # Medido en dark-kitchen: «PREPARACIÓN MAÑANA», tareas=4, casadas=1,
    # ratio=0,25, dos veces, idénticas salvo la fila. La clave pasa a ser
    # (hoja SIN nº de fila, banda, tareas, casadas, ratio) y se conserva la
    # ÚLTIMA medida, que es la que se tomó sobre la geometría FINAL del libro
    # —la que de verdad se entrega—: quedarse con la primera citaría una fila
    # en la que ya no está la banda.
    vistos, unicos = {}, []
    for d in motor.SOLAPES:
        clave = (d['hoja'].split('!')[0], d['banda'], d['tareas'],
                 d['casadas'], d['ratio'])
        # `destinos` y `anotada` NO entran en la clave (los fija la misma
        # medición que da `casadas` y `ratio`) pero sí se vigilan: si dos
        # mediciones de la misma banda discrepasen en ellos, se publican las
        # dos, que es el criterio de T-05.
        firma = (tuple(d['destinos']), d['anotada'])
        anterior = vistos.get(clave)
        if anterior is not None and anterior[0] == firma:
            unicos[anterior[1]] = d              # se queda la ÚLTIMA medida
            continue
        vistos[clave] = (firma, len(unicos))
        unicos.append(d)
    solapes = sorted(unicos, key=lambda d: -d['ratio'])
    if solapes:
        log('  solape medido banda↔marco (umbral de anotación '
            f'{motor.UMBRAL_BANDA:.0%}): '
            + ' · '.join(f"{d['hoja']} {d['ratio']:.0%}"
                         + ('' if d['anotada'] else ' (NO anotada)')
                         for d in solapes[:6]))

    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 3/7 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(idem_dir):
            shutil.rmtree(idem_dir)
        shutil.copytree(carpeta, idem_dir)
        antes = {n: digest(os.path.join(carpeta, n)) for n in nombres}
        motor.contexto(
            idem_dir, nombres,
            lambda f: openpyxl.load_workbook(os.path.join(idem_dir, f)),
            producto=pid)
        motor.CTX['producto'] = pid
        for fname in nombres:
            procesar(idem_dir, fname, etapas, contenido, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(idem_dir, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:30]}
        log(f'  diferencias 1.ª vs 2.ª pasada: {len(difs)}')
        for d in difs[:10]:
            log('    ' + d)
        motor.contexto(
            carpeta, nombres,
            lambda f: openpyxl.load_workbook(os.path.join(carpeta, f)),
            producto=pid)
        motor.CTX['producto'] = pid

    log('\n== 4/7 · inject_cache (al final del todo) ==')
    guardados = [f['fichero'] for f in informe_ficheros if f['guardado']]
    cache = inject_cache(carpeta, guardados)

    log('\n== 5/7 · data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log(f"  con valor: {ver['con_valor']} · fallos: {len(ver['fallos'])}")
    for f in ver['fallos'][:8]:
        log('    ' + f)

    log('\n== 6/7 · pycel con los casos de §6 + DV y bio ==')
    demostraciones = []
    if not args.sin_demos:
        if os.path.isdir(demos_dir):
            shutil.rmtree(demos_dir)
        demostraciones = (demo_arqueo(carpeta, demos_dir)
                          + demo_contador(carpeta, demos_dir)
                          + demo_07(carpeta, demos_dir)
                          + demo_p4(carpeta, nombres, demos_dir))
    for c in demostraciones:
        log(f"  {c['caso']} · {c['ref']}: {c['obtenido']!r} "
            f"{'OK' if c['ok'] else 'FALLA (esperaba ' + repr(c['esperado']) + ')'}")
    gates = gate_dv_y_bio(carpeta, nombres)
    con_ins = len(nombres) - len(gates['sin_hoja_instrucciones'])
    log(f"  DV «✓,—,N/A»: {gates['hojas_checklist']} hojas de checklist, "
        f"{len(gates['dv_incorrectas'])} incorrectas · bio en "
        f"{gates['instrucciones_con_bio']}/{con_ins} Instrucciones del "
        f"producto · {len(gates['version_desfasada'])} sin la versión 2.0")
    for d in gates['dv_incorrectas'][:8]:
        log('    ' + d)
    for d in gates['version_desfasada'][:8]:
        log('    ' + d)
    for d in gates['sin_hoja_instrucciones']:
        log(f'    AVISO (j) · {d}: no tiene hoja «Instrucciones» — no se '
            'inventa ninguna; lo decide el orquestador')
    for m in gates['aviso_dv_mezcladas']:
        log('    AVISO TEC-R2-08 · el producto tiene MÁS de una lista de '
            'desplegable: ' + m)
    prec = gate_precargado(carpeta)
    if prec.get('sin_caja'):
        # CB-E7 — informativo, no fallo: sin fichero de negocio no hay nada que
        # precargar. Se imprime la razón, no un «fichero None … OK» que leería
        # como un gate aprobado sobre un sujeto que no existe.
        for d in prec.get('informativo', []):
            log('  08 precargado: INFORMATIVO · ' + d)
    else:
        log(f"  08 precargado: fichero {prec['fichero']} "
            f"(firma de negocio {'OK' if prec.get('firma_ok') else 'NO VÁLIDA'})"
            f" · {prec['tareas']} tareas, {len(prec['huecos'])} sin "
            'Responsable u Hora')
    for h in prec.get('huecos_firma', []):
        log('    ' + h)
    for h in prec['huecos'][:6]:
        log('    ' + h)
    rec = gate_recuento(carpeta, nombres)
    log(f"  recuento de tareas entregadas: {rec['total']} en total "
        f"(molde ▸ {rec['recuento_por_molde']['checklist_▸']} + molde P4 "
        f"{rec['recuento_por_molde']['p4']}) · "
        + ' · '.join(f'{k.split("-")[0]}={v}'
                     for k, v in rec['por_fichero'].items() if v))
    meta = gate_metadata(carpeta, nombres)
    log(f"  metadata (m5): subject «{meta['subject_esperado']}» · keywords "
        f"«{meta['keywords_esperado']}» · {len(meta['incoherentes'])} "
        f"incoherencias en {len(nombres)} ficheros")
    for d in meta['incoherentes'][:8]:
        log('    ' + d)
    for d in meta['keywords_propias'][:4]:
        log('    AVISO m5 · ' + d)
    tpv = gate_tpv(carpeta, nombres)
    log(f"  TPV: {len(tpv['tareas_encender_tpv'])} tareas «Encender …TPV» en "
        f"el producto ({len(tpv['fuera_del_fichero_de_negocio'])} fuera del "
        'fichero de negocio)')
    for d in tpv['tareas_encender_tpv']:
        log(f"    {d['ref']} [{d['papel']}]: {d['texto']}")

    orto = gate_ortografia(carpeta, nombres)
    log('  CB-E1 ortografía: ' + orto['resumen']
        + ('' if orto['aplica'] else ' (el paso NO aplica: el producto no es '
                                    'de la sub-familia ChefBusiness)'))
    for d in orto['lemas_vivos'][:8]:
        log(f"    {d['ref']}: {d['texto']}")
    cont = gate_contadores(carpeta, nombres)
    log(f"  CB-E2 contadores: {cont['con_contador']} hojas de checklist con "
        f"fila de totales, {len(cont['sin_contador'])} sin ella")
    for d in cont['sin_contador']:
        log('    ' + d)
    moldes = gate_moldes(carpeta, nombres)
    log(f"  molde: {len(moldes['hojas_sin_molde'])} hojas sin clasificar")
    for d in moldes['hojas_sin_molde'][:10]:
        log('    AVISO molde · ' + d + ': ningún molde la reconoce (no recibe '
            'DV, ni contador, ni A4, ni protección)')

    log('\n== 7/7 · censo-entregables --fail ==')
    cen = censo(carpeta)

    fallos = []
    if idem.get('diferencias'):
        fallos.append(f"idempotencia: {idem['diferencias']} diferencias")
    if ver['fallos']:
        fallos.append(f"cache: {len(ver['fallos'])} fórmulas sin valor")
    fallos += [f"§6 {c['caso']}: esperaba {c['esperado']!r}, dio "
               f"{c['obtenido']!r}" for c in demostraciones if not c['ok']]
    fallos += gates['dv_incorrectas']
    fallos += prec.get('huecos_firma', [])
    fallos += prec['huecos'][:10]
    # (j) — la bio y la versión se exigen en TODO fichero del producto que
    # tenga hoja «Instrucciones», esté o no en alcance del molde ▸.
    fallos += [f'sin bio en Instrucciones: {n}' for n in gates['sin_bio']]
    fallos += gates['version_desfasada']
    # m5 — la metadata del producto entero (también la de los P4 y los BONUS)
    fallos += meta['incoherentes']
    fallos += [f'{n}: en alcance y SIN hoja «Instrucciones»'
               for n in gates['sin_hoja_instrucciones']
               if n in ctx['ficheros']]
    # CB-E1 §1.3 — rojo si sobrevive UN solo lema del diccionario.
    fallos += [f"ortografía: {d['ref']}: {d['texto']}"
               for d in orto['lemas_vivos'][:20]]
    # CB-E2 §1.3 — rojo si una hoja de checklist se queda sin contador.
    fallos += cont['sin_contador']
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió ' + str(cen['exit']))
    if not args.sin_demos and not demostraciones:
        fallos.append('§6: no se pudo ejecutar ninguna demostración')

    informe = {
        'producto': pid, 'version': '2.0', 'spec': SPEC,
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'etapas': sorted(etapas),
        'carpeta_origen': origen, 'carpeta_trabajo': carpeta,
        # m6/m8 — queda escrito en el informe qué encontró el paso del 09 de
        # catering (el 09 viejo o el nuevo ya renombrado) y qué hizo. Sin esto
        # el informe no distinguiría una pasada antes del `git mv` de una
        # después, y las dos producen el mismo fichero final.
        'sustitucion_09_catering': sust09,
        'contexto': {k: (sorted(v) if isinstance(v, set) else v)
                     for k, v in ctx.items()},
        'ficheros': informe_ficheros,
        'fuera_de_alcance': [f['fichero'] for f in informe_ficheros
                             if not f['en_alcance']],
        'inventario': inventario(carpeta, nombres),
        'gates': {
            'idempotencia': idem,
            'inject_cache': cache,
            'data_only_formulas_nuevas': ver,
            'dv_y_bio': gates,
            'negocio_precargado': prec,
            'recuento_tareas': rec,
            'metadata': meta,
            'tpv_duplicado': tpv,
            'bandas_solapadas': solapes,
            'censo_entregables': cen,
            'ortografia': orto,
            'contadores': cont,
            'molde': moldes,
        },
        'demostraciones_spec_6': demostraciones,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1)
        log(f'\ninforme → {args.json}')

    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'TODO VERDE (idempotencia, cache, §6, DV, bio, censo)'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
