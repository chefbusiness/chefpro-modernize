#!/usr/bin/env python3
"""
motor.py — Motor común del Kit Plan Financiero para Restaurantes v2.0.

Implementa el **§1** de `kit-plan-financiero-v2-SPEC.md` sobre los **10** xlsx de
`astro-site/public/dl/kit-plan-financiero/` (47 hojas). NO toca ficheros: recibe
un `Workbook` ya cargado y lo modifica en memoria; quien guarda es `main.py`.

Qué cubre (§1.1-§1.8):
  1.1  Gráficos de verdad con `openpyxl.chart` en 9 de los 10 ficheros
       (`graficos()`), con **precondición**: si el bloque de datos que el
       gráfico necesita todavía no existe —porque lo crea un grupo que aún no
       está construido— el gráfico se OMITE y queda anotado como `pendiente`,
       nunca se dibuja sobre celdas vacías.
  1.2  Formato condicional real (`semaforo()`, `regla_expresion()`): hoy hay
       **0 reglas** en las 47 hojas y tres Instrucciones prometen colores.
  1.3  `IFERROR` en las divisiones desnudas (`guardas()`), protección de hoja
       SIN contraseña con los verdes desbloqueados (`proteger()`) y validación
       de datos ≥ 0 en importes / 0-1 en porcentajes (`validaciones()`).
  1.4  Datos de ejemplo etiquetados: fila «VALORES DE EJEMPLO …» en cada hoja
       de datos precargada (`fila_ejemplo()`), y la base numérica común del kit
       en `EJEMPLO_CANONICO` — la usan grupo_b (02) y grupo_c (06, BONUS-08)
       para que los tres ejemplos cuenten la MISMA historia (§4 BONUS-08).
  1.5  Base «sin IVA» declarada (`sin_iva()`): «(sin IVA)» en cada etiqueta de
       ventas y de ticket de 01, 01b, 02, 05, 06 y 07 + la línea que explica
       que el 03 va CON IVA porque es caja.
  1.6  Formatos, anchos y cabeceras (`formatos()`, `anchos()`): incluye crear
       `07!'Resumen Ejecutivo'!C5:C25`, la única columna de valores del kit sin
       celdas ni verde.
  1.7  Instrucciones, **bio anclada** (INSERCIÓN: no la lleva ninguno de los
       10) y línea de versión 2.0 (`cierre_instrucciones()`), más metadata
       `title`/`subject`/`keywords` → `… · v2.0` (`metadatos()`).
  1.8  Orden de construcción y **caché de la TIR** (`cachear_irr()`): pycel no
       implementa `IRR` y `IFERROR` no lo atrapa, así que `07!Proyecciones!B21`
       es la única fórmula del kit que `inject_cache.py` deja sin `<v>`. Se
       calcula por Newton-Raphson y se inyecta con el mismo mecanismo de zip.

CONVENCIONES DE FAMILIA que este motor garantiza (`cerrar()`):
  · editables **verdes E8F5E9 y desbloqueadas**, calculadas sin relleno;
  · `IFERROR` en toda división; parámetros en celda, nunca literales;
  · protección de hoja **sin contraseña**;  · A4 completo (paperSize 9 +
    fitToPage + pie) que el censo exige;  · ninguna celda con `''` (el censo la
    cuenta como defecto `empty_str`);  · `creator = 'AI Chef Pro'`.

IDEMPOTENCIA: todo lo que escribe el motor es **absoluto** (mismo valor en la
misma celda). Los tres puntos con estado —formato condicional, validaciones y
gráficos— se PURGAN antes de reescribirse, que es justo donde openpyxl acumula
duplicados en una segunda pasada. La única escritura que depende de dónde estaba
algo (el bloque de cierre de Instrucciones) se ancla borrando primero las líneas
del bloque y recalculando el final del texto, no leyendo la posición anterior.

pycel 1.0b30 (medido, SPEC cabecera): implementa `NPV`, `ROUNDUP`, `ROUND`,
`MAX`, `MIN`, `ABS`, `AVERAGE`, `SUM`, `SUMIF`, `COUNTIF`, `INDEX`, `MATCH`,
`IFERROR`, `TEXT` y `^`; **NO** `IRR`, `PMT` ni `COUNTA`. Sustitutos
obligatorios: `COUNTA(r)` → `COUNTIF(r,"<>")`; `PMT` → anualidad algebraica
`importe*i/(1-(1+i)^-n)`; `IRR` → fórmula en la celda + `cachear_irr()`.

Utilidades `_traducir_formula`, `_rangos_dv`, `_restaurar_dv`,
`_desplazar_rango`, `insertar_columna`, `insertar_fila` y `linea_instrucciones`
copiadas del motor del Kit de Escandallos v2.0 (probado en producción el
2026-08-22); se copian y no se importan para no arrastrar su `main()`.
"""
import copy
import html
import re
import shutil
import zipfile

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Paleta, formatos y textos fijos
# ==========================================================================
VERDE = 'E8F5E9'          # celda editable (verde de edición de la familia)
GRIS = 'F2F2F2'           # bloque auxiliar «datos base, no son caja»
CREMA = 'FFF8DC'          # subtotal
CAB = '2D2D2D'            # cabecera de tabla (ya en v1.1)

# Semáforo por FORMATO CONDICIONAL (§1.2). Los de `Benchmarks` son otros
# (C8E6C9/FFF9C4/FFCDD2, relleno fijo escrito en v1.1) y no se tocan.
CF_VERDE_BG, CF_VERDE_FG = 'C6EFCE', '006100'
CF_AMBAR_BG, CF_AMBAR_FG = 'FFEB9C', '9C6500'
CF_ROJO_BG, CF_ROJO_FG = 'FFC7CE', '9C0006'

FMT_EUR = '#,##0.00 €'
FMT_PCT = '0.0%'
FMT_PCT0 = '0%'
FMT_ENT = '#,##0'
FMT_ENT0 = '0'
FMT_FECHA = 'dd/mm/yyyy'
FMT_X = '0.00"x"'
FMT_PP = '0.0" p.p."'

PRODUCTO = 'Kit Plan Financiero para Restaurantes'
SUBJECT_V2 = PRODUCTO + ' · v2.0'
KEYWORDS_V2 = ('kit plan financiero, plan de negocio restaurante, viabilidad, '
               'punto de equilibrio, cash flow, TIR VAN payback, AI Chef Pro')

VERSION_LINE = ('Versión 2.0 · agosto 2026 · aichef.pro/kit-plan-financiero · '
                'info@aichef.pro')
RX_VERSION = re.compile(r'^Versi[óo]n \d+\.\d+ · ')

# §1.7 — la bio NO la lleva ninguno de los 10: aquí es INSERCIÓN, no
# sustitución. `RX_BIO` sirve sólo para no duplicarla en la 2.ª pasada y para
# reconocer una redacción previa distinta.
BIO_LINE = ('Diseñado por John Guerrero — chef y consultor gastronómico desde '
            '2010, en cocina desde los 17 años · johnguerrero.es')
RX_BIO = re.compile(r'John\s+Guerrero')

# §1.5 — la base de las cifras, dicha en voz alta en cada libro.
NOTA_IVA = ('Todas las cifras van SIN IVA; en el 03 (tesorería) van CON IVA '
            'porque es caja.')
NOTA_IVA_CAJA = ('Todas las cifras de esta plantilla van CON IVA: es caja. '
                 'En el resto del kit (01, 01b, 02, 05, 06 y 07) van SIN IVA.')
RX_NOTA_IVA = re.compile(r'^Todas las cifras')

NOTA_DESPROTEGER = ('Para editar una celda que no esté en verde: Revisar → '
                    'Desproteger hoja (no tiene contraseña).')
RX_DESPROTEGER = re.compile(r'^Para editar una celda')

EJEMPLO_LINEA = 'VALORES DE EJEMPLO — sustitúyelos por los tuyos'
RX_EJEMPLO = re.compile(r'^VALORES DE EJEMPLO')

FICHEROS = [
    '01-plan-financiero-previsional.xlsx',
    '01b-plan-financiero-previsional-5-anos.xlsx',
    '02-calculadora-punto-equilibrio.xlsx',
    '03-cash-flow-forecast.xlsx',
    '04-presupuesto-inversion-capex.xlsx',
    '05-pyl-mensual-real-vs-presupuesto.xlsx',
    '06-dashboard-ratios-financieros.xlsx',
    '07-informe-viabilidad-bancos.xlsx',
    'BONUS-08-simulador-escenarios.xlsx',
    'BONUS-09-checklist-pre-apertura.xlsx',
]

# El 03 es el único libro cuya base es CAJA (con IVA): §1.5.
FICHERO_CAJA = '03-cash-flow-forecast.xlsx'

# Ficheros donde la línea «Todas las cifras van SIN IVA» NO tiene sentido
# (RC-28): la checklist no tiene importes y el CAPEX desglosa base + IVA.
SIN_NOTA_IVA = ('BONUS-09-checklist-pre-apertura.xlsx',
                '04-presupuesto-inversion-capex.xlsx')

# ==========================================================================
# Taxonomía canónica — las 10 categorías del kit
# ==========================================================================
# «Coherentes entre sí: mismas líneas de ingreso y gasto … en las 10
# plantillas» (§5, sustituye a las «fórmulas encadenadas entre libros» que no
# existen). Son EXACTAMENTE las 4 líneas de ingreso + 6 de gasto que ya usan
# 01!'Año 1'!A6:A9 / A13:A18 y 05!Ene!A6:A9 / A13:A18; los alias son las
# variantes con que la misma línea aparece en 02!Datos, 03!'Flujo Mensual',
# 06!Ratios y 07!Proyecciones. Cualquier grupo que cree un desplegable de
# categoría, una fila nueva o una columna de IVA lee de aquí: si cada libro se
# inventa su lista, vuelve el problema que la FAQ describe al revés.
#
# `iva` es el tipo que le corresponde a esa línea en España (§3 del 03 y §3 del
# 04): 10 % en restauración —repercutido en las ventas y soportado en las
# compras de alimentación— y 21 % en el resto de gastos y en el CAPEX.
IVA_RESTAURACION = 0.10
IVA_GENERAL = 0.21

TAXONOMIA = (
    # (clave, etiqueta canónica, tipo, iva, (alias…))
    ('comedor', 'Comedor', 'ingreso', IVA_RESTAURACION,
     ('Ventas Comedor', 'Ventas comedor (IVA incl.)', 'Ventas comedor')),
    ('barra', 'Barra', 'ingreso', IVA_RESTAURACION,
     ('Ventas Barra', 'Ventas barra (IVA incl.)', 'Ventas barra')),
    ('delivery', 'Delivery', 'ingreso', IVA_RESTAURACION,
     ('Ventas Delivery', 'Ventas delivery (IVA incl.)', 'Ventas delivery')),
    ('eventos', 'Eventos / Catering', 'ingreso', IVA_RESTAURACION,
     ('Eventos / Catering', 'Eventos y catering')),
    ('food', 'Food Cost', 'gasto', IVA_RESTAURACION,
     ('Coste materia prima (€)', 'Compras materia prima', 'Food Cost (€)')),
    ('personal', 'Personal (Labor)', 'gasto', 0.0,
     ('Personal', 'Nóminas + SS', 'Coste personal total (€)',
      'Labor Cost (€)', 'Nóminas fijas (bruto)')),
    ('alquiler', 'Alquiler', 'gasto', IVA_GENERAL,
     ('Alquiler mensual',)),
    ('suministros', 'Suministros', 'gasto', IVA_GENERAL,
     ('Suministros (luz, agua, gas)',)),
    ('marketing', 'Marketing', 'gasto', IVA_GENERAL,
     ('Marketing fijo',)),
    ('admin', 'Administración / Otros', 'gasto', IVA_GENERAL,
     ('Gestoría / Admin', 'Gestoría / Administración', 'Otros gastos fijos')),
)

CATEGORIAS = [t[1] for t in TAXONOMIA]
CAT_INGRESO = [t[1] for t in TAXONOMIA if t[2] == 'ingreso']
CAT_GASTO = [t[1] for t in TAXONOMIA if t[2] == 'gasto']
IVA_POR_CATEGORIA = dict((t[1], t[3]) for t in TAXONOMIA)


def categoria_de(etiqueta):
    """Etiqueta (canónica o alias, con o sin sufijo «(sin IVA)») → clave."""
    if not isinstance(etiqueta, str):
        return None
    base = etiqueta.split(' (')[0].strip()
    for clave, canon, _tipo, _iva, alias in TAXONOMIA:
        if base == canon.split(' (')[0] or etiqueta in alias or base in alias:
            return clave
    return None


def iva_de(etiqueta):
    """IVA que le toca a una línea del kit; `None` si no es del catálogo."""
    clave = categoria_de(etiqueta)
    if clave is None:
        return None
    for c, canon, _t, iva, _a in TAXONOMIA:
        if c == clave:
            return iva
    return None


# Unidades del kit: son las de los parámetros, no las de un almacén. Sirven
# para el desplegable de «Período» y para que ningún grupo escriba «mensual»
# con minúscula en un libro y «Mensual» en otro.
UNIDADES_PERIODO = ('Mensual', 'Trimestral', 'Anual')
UNIDADES_PLAZO = ('Días', 'Meses', 'Años')

# §1.4 / §7.5 — la base numérica común. El ejemplo del 02 describía hoy un
# negocio que el propio kit califica de inviable (labor cost 41 %, prime cost
# ~71 %) y el BONUS-08 lo empeoraba contando el personal dos veces. Estos son
# los números que 02, 06 y BONUS-08 tienen que contar a la vez.
EJEMPLO_CANONICO = {
    'ticket': 22.0,               # € por comensal, sin IVA
    'cubiertos_dia': 55,          # cubiertos/día (NO «ocupación»: COM-26)
    'dias_apertura': 26,          # días/mes
    'ventas_mes': 22.0 * 55 * 26,                     # 31.460,00 €
    'food_cost_pct': 0.30,
    'labor_cost_pct': 0.28,
    # Alquiler 3.000 + seguros 400 + suministros 1.500 + marketing 500 +
    # gestoría 300 + otros 500 = 6.200 €. Son EXACTAMENTE los del 02 sin
    # las nóminas ni la cuota: si no coinciden, el 02 y el BONUS-08 dan
    # EBITDA distinto para el mismo restaurante.
    'fijos_sin_personal': 6200.0,
    'cuota_prestamo': 800.0,        # fuera del EBITDA (TEC-23)
    'aforo_plazas': 60,             # plazas para el RevPASH (DOM-07)
    # Horas de SERVICIO, no de apertura: el RevPASH mide plaza·hora con
    # clientes sentados. 6 h/día × 26 días = 156. Con las 360 «horas de
    # servicio/mes» de la v1.1 (12 h/día × 30) el ratio salía cuatro veces más
    # bajo de lo que le toca y el benchmark parecía inalcanzable.
    'horas_servicio_mes': 156,
    'm2_sala': 80,
}

# ==========================================================================
# Vocabularios de semáforo (§1.2)
# ==========================================================================
SEM_DESVIACION = (('🔴 Alerta', CF_ROJO_BG, CF_ROJO_FG),
                  ('⚠️ Atención', CF_AMBAR_BG, CF_AMBAR_FG),
                  ('✅ OK', CF_VERDE_BG, CF_VERDE_FG))
SEM_RATIO = (('🔴 Alto', CF_ROJO_BG, CF_ROJO_FG),
             ('⚠️ Aceptable', CF_AMBAR_BG, CF_AMBAR_FG),
             ('✅ Excelente', CF_VERDE_BG, CF_VERDE_FG))
SEM_GOP = (('🔴 Peligro', CF_ROJO_BG, CF_ROJO_FG),
           ('⚠️ Ajustado', CF_AMBAR_BG, CF_AMBAR_FG),
           ('✅ Sano', CF_VERDE_BG, CF_VERDE_FG))
SEM_TESORERIA = (('BAJO MÍNIMO', CF_ROJO_BG, CF_ROJO_FG),
                 ('OK', CF_VERDE_BG, CF_VERDE_FG))
SEM_BANCO = (('🔴', CF_ROJO_BG, CF_ROJO_FG),
             ('⚠️', CF_AMBAR_BG, CF_AMBAR_FG),
             ('✅', CF_VERDE_BG, CF_VERDE_FG))
SEM_ESTADO = (('Completada', CF_VERDE_BG, CF_VERDE_FG),
              ('En curso', CF_AMBAR_BG, CF_AMBAR_FG),
              ('Pendiente', CF_ROJO_BG, CF_ROJO_FG))

# ==========================================================================
# Registro de fórmulas (main.py verifica una por una que quedaron cacheadas)
# ==========================================================================
REGISTRO = []


def reg(ws, coord, formula):
    REGISTRO.append((ws.title, coord, formula))


def f(ws, coord, formula, fmt=None, align=None):
    """Escribe una FÓRMULA y la registra para la verificación `data_only`."""
    cel = ws[coord]
    cel.value = formula
    if fmt:
        cel.number_format = fmt
    if align:
        cel.alignment = Alignment(horizontal=align)
    reg(ws, coord, formula)
    return cel


def val(ws, coord, valor, fmt=None, verde_=False, bold=None, align=None):
    """Escribe un VALOR (constante). `verde_` lo marca como editable."""
    cel = ws[coord]
    cel.value = valor
    if fmt:
        cel.number_format = fmt
    if verde_:
        cel.fill = PatternFill('solid', fgColor=VERDE)
        cel.protection = Protection(locked=False)
    if bold is not None:
        cel.font = Font(bold=bold, size=cel.font.size, color=cel.font.color)
    if align:
        cel.alignment = Alignment(horizontal=align)
    return cel


def iferror(expresion, alterna=''):
    """`=IFERROR(<expr>,"<alterna>")` — §1.3, guarda en TODA división.

    `expresion` va SIN el `=` inicial. `alterna` vacía produce `""`, que es lo
    que el motor cuenta como «vacía por diseño» al verificar el caché.
    """
    expr = expresion[1:] if expresion.startswith('=') else expresion
    return '=IFERROR(' + expr + ',"' + alterna + '")'


def es_verde(cel):
    relleno = cel.fill
    return (relleno is not None and relleno.fill_type == 'solid'
            and relleno.fgColor is not None
            and isinstance(relleno.fgColor.rgb, str)
            and relleno.fgColor.rgb.upper().endswith(VERDE))


def verde(ws, rango):
    """Marca en verde (y desbloquea) un rango 'A1:C3' o una celda 'A1'."""
    for fila in ws[rango] if ':' in rango else [[ws[rango]]]:
        for cel in fila:
            cel.fill = PatternFill('solid', fgColor=VERDE)
            cel.protection = Protection(locked=False)


def limpiar_rango(ws, rango):
    """Vacía un rango antes de reescribirlo con un mapa de filas NUEVO.

    No basta con poner `value = None`: si el relleno verde de la v1.1 sobrevive
    en una celda que ya no es un input, `proteger()` la deja desbloqueada y
    `validaciones()` le cuelga una DV que no le corresponde. Se limpian valor,
    relleno, formato de número y bloqueo.
    """
    # Las combinaciones se DESHACEN primero: en estos diez ficheros el pie
    # «© 2026 AI Chef Pro» va combinado a lo ancho de la tabla, y una
    # `MergedCell` tiene el `value` de sólo lectura — asignarle nada revienta
    # con AttributeError. Se rehace el pie donde toque, sin combinar.
    objetivo = CellRange(rango) if ':' in rango else CellRange(rango + ':'
                                                              + rango)
    for m in [str(r) for r in ws.merged_cells.ranges]:
        # `intersection` LANZA ValueError cuando no hay solape (no devuelve
        # None): hay que atraparlo, no comprobarlo.
        try:
            CellRange(m).intersection(objetivo)
        except ValueError:
            continue
        ws.unmerge_cells(m)
    for fila in ws[rango] if ':' in rango else [[ws[rango]]]:
        for cel in fila:
            cel.value = None
            cel.fill = PatternFill()
            cel.number_format = 'General'
            cel.protection = Protection(locked=True)


def aplicar_estilo(ws, coord, estilo, fmt=None):
    """Copia un estilo y REPONE el formato de número.

    `_style` arrastra el `numFmtId`, así que asignarlo DESPUÉS de `f()`/`val()`
    pisaba en silencio el formato que se acababa de poner. Se veía sólo en la
    idempotencia: `0.0%` en la primera pasada y `General` en la segunda.
    """
    ws[coord]._style = copy.copy(estilo)
    if fmt:
        ws[coord].number_format = fmt
    return ws[coord]


def marcar_editable(ws, rango):
    """Desbloquea por ROL, no por color (RD-06).

    `proteger()` desbloqueaba SÓLO lo pintado de `E8F5E9`. En `02!Escenarios`
    los inputs van pintados con los colores de escenario (rojo/ámbar/verde de
    semáforo), así que la hoja entera quedaba con CERO celdas editables: un
    lunes no se podía cambiar ni el ticket ni los cubiertos. Aquí se declara
    que una celda es input aunque su relleno diga otra cosa.
    """
    marcadas = getattr(ws, '_pf_editables', None)
    if marcadas is None:
        marcadas = set()
        ws._pf_editables = marcadas
    for fila in ws[rango] if ':' in rango else [[ws[rango]]]:
        for cel in fila:
            marcadas.add(cel.coordinate)
            cel.protection = Protection(locked=False)


def permitir_negativo(ws, rango):
    """Excluye un rango de la validación «≥ 0» del §1.3 (RD-30/RT-22/RT-07).

    Hay celdas que TIENEN que poder ser negativas: el saldo inicial de una
    tesorería en descubierto —el caso que el 03 existe para detectar—, un
    EBITDA presupuestado o el flujo del año 0. Con la DV genérica, Excel
    rechazaba el dato correcto con un error bloqueante y el usuario acababa
    desprotegiendo la hoja el primer día, que es justo lo que la protección
    venía a evitar.
    """
    libres = getattr(ws, '_pf_negativos', None)
    if libres is None:
        libres = set()
        ws._pf_negativos = libres
    for fila in ws[rango] if ':' in rango else [[ws[rango]]]:
        for cel in fila:
            libres.add(cel.coordinate)


def gris(ws, rango):
    """Bloque auxiliar «datos base, no son caja» (§3 del 03)."""
    for fila in ws[rango] if ':' in rango else [[ws[rango]]]:
        for cel in fila:
            cel.fill = PatternFill('solid', fgColor=GRIS)


# ==========================================================================
# Formato condicional (§1.2) — SIEMPRE purgando antes de escribir
# ==========================================================================
def _norm_rango(ref):
    try:
        return CellRange(ref).coord
    except Exception:                                        # noqa: BLE001
        return ref


def _limpiar_cf(ws, rango):
    """Quita las reglas cuyo sqref es exactamente `rango`.

    Sin esto la 2.ª pasada apila una copia de cada regla sobre el mismo rango:
    el fichero sigue viéndose igual en Excel, pero la idempotencia salta y el
    XML crece en cada ejecución.
    """
    # El sqref se compara NORMALIZADO: openpyxl guarda `C5:C5` como `C5`, así
    # que comparar cadenas sueltas no encontraba la regla anterior y la 2.ª
    # pasada apilaba una copia (idempotencia rota por una sola regla).
    objetivo = _norm_rango(rango)
    supervivientes = []
    for cf in ws.conditional_formatting:
        if _norm_rango(str(cf.sqref)) == objetivo:
            continue
        supervivientes.append((str(cf.sqref), list(cf.rules)))
    nueva = ConditionalFormattingList()
    for sqref, reglas in supervivientes:
        for r in reglas:
            nueva.add(sqref, r)
    ws.conditional_formatting = nueva


def _dxf(bg, fg):
    return DifferentialStyle(font=Font(color=fg, bold=True),
                             fill=PatternFill(start_color=bg, end_color=bg,
                                              fill_type='solid'))


def semaforo(ws, rango, vocabulario=SEM_DESVIACION,
             exigir_datos=True):
    """Colorea por TEXTO contenido: verde/ámbar/rojo de verdad.

    Hoy el kit tiene 0 reglas de formato condicional y los emojis salen en
    negro: en el A4 en blanco y negro que la Fase A dejó listo para imprimir,
    «🔴 Alerta» y «✅ OK» son el mismo gris. Las reglas se añaden de MÁS grave
    a menos y con `stopIfTrue`, porque «✅ OK» contiene «OK» y si ganara la
    regla verde, «⚠️ BAJO MÍNIMO» no se pintaría.
    """
    _limpiar_cf(ws, rango)
    ancla = rango.split(':')[0]
    # RT-17: una regla sobre un rango SIN una sola celda con contenido no puede
    # encenderse jamás, y el gate contaba reglas, no efectos. Si no hay ancla
    # con valor, no se escribe la regla y se devuelve False para que quien
    # llama lo anote como pendiente.
    if exigir_datos:
        hay = False
        for fila in ws[rango] if ':' in rango else [[ws[rango]]]:
            for cel in fila:
                if cel.value is not None:
                    hay = True
                    break
            if hay:
                break
        if not hay:
            return False
    for texto, bg, fg in vocabulario:
        regla = Rule(type='containsText', operator='containsText', text=texto,
                     dxf=_dxf(bg, fg), stopIfTrue=True)
        regla.formula = ['NOT(ISERROR(SEARCH("' + texto + '",' + ancla + ')))']
        ws.conditional_formatting.add(rango, regla)
    return True


def regla_expresion(ws, rango, formula, bg=CF_ROJO_BG, fg=CF_ROJO_FG,
                    parar=True):
    """Regla `expression` (p. ej. `=$C6<$C$3` para el saldo bajo umbral)."""
    _limpiar_cf(ws, rango)
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=parar,
                           font=Font(color=fg, bold=True),
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type='solid')))


# ==========================================================================
# Validación de datos (§1.3) — también purgando
# ==========================================================================
def _purgar_dv(ws, sqref, tipo):
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if not (dv.type == tipo and str(dv.sqref) == sqref)]


def dv_lista(ws, rango, opciones, titulo='Valor no válido', mensaje=None):
    """Desplegable con `showErrorMessage=True` (§1.3)."""
    formula = '"' + ','.join(opciones) + '"'
    _purgar_dv(ws, rango, 'list')
    dv = DataValidation(type='list', formula1=formula, allow_blank=True,
                        showErrorMessage=True, errorTitle=titulo,
                        error=mensaje or ('Elige un valor de la lista: '
                                          + ', '.join(opciones)))
    ws.add_data_validation(dv)
    dv.add(rango)
    return dv


def dv_numerica(ws, coordenadas, minimo=0, maximo=None, titulo=None,
                mensaje=None):
    """Una sola DV para un conjunto de celdas (se pasa la lista ordenada).

    Con `maximo=None` es «≥ minimo» (importes); con `maximo=1` es el 0-1 de
    los porcentajes, que en Excel se teclean como 0,35 y no como 35.
    """
    if not coordenadas:
        return None
    coords = sorted(set(coordenadas),
                    key=lambda c: (column_index_from_string(
                        re.match(r'([A-Z]+)', c).group(1)),
                        int(re.search(r'(\d+)', c).group(1))))
    if maximo is None:
        dv = DataValidation(type='decimal', operator='greaterThanOrEqual',
                            formula1=str(minimo), allow_blank=True,
                            showErrorMessage=True,
                            errorTitle=titulo or 'Importe no válido',
                            error=mensaje or 'Escribe un número mayor o igual '
                            'que ' + str(minimo) + '.')
    else:
        dv = DataValidation(type='decimal', operator='between',
                            formula1=str(minimo), formula2=str(maximo),
                            allow_blank=True, showErrorMessage=True,
                            errorTitle=titulo or 'Porcentaje no válido',
                            error=mensaje or 'Escribe un porcentaje entre '
                            + str(minimo) + ' y ' + str(maximo)
                            + ' (0,35 = 35 %).')
    ws.add_data_validation(dv)
    for c in coords:
        dv.add(c)
    _purgar_dv_duplicadas(ws)
    return dv


def _purgar_dv_duplicadas(ws):
    """Deja una sola DV por (tipo, formula1, formula2, sqref).

    La 2.ª pasada vuelve a crear las mismas validaciones sobre las mismas
    celdas; sin esto, el `.xlsx` acumula validaciones idénticas y el digest de
    idempotencia las ve.
    """
    vistas, fuera = set(), []
    for dv in ws.data_validations.dataValidation:
        clave = (dv.type, dv.formula1, dv.formula2, str(dv.sqref))
        if clave in vistas:
            continue
        vistas.add(clave)
        fuera.append(dv)
    ws.data_validations.dataValidation = fuera


# §1.3 — celdas que la SPEC define como negativas o que pueden serlo por
# naturaleza. Las pone el motor incluso si el grupo que reescribe la hoja no
# está cargado (RD-30/RT-22).
NEGATIVOS = {
    '03-cash-flow-forecast.xlsx': {'Flujo Mensual': ['B5']},
}


def validaciones(ws, informe=None, fname=None):
    """§1.3: DV numérica sobre TODAS las celdas verdes de la hoja.

    Se clasifica por formato de número: `€`/`#,##0` → importe ≥ 0;
    `%` → 0-1. Las verdes de texto o de fecha se quedan sin validación (un
    nombre de restaurante no tiene rango) y las que ya tienen desplegable
    tampoco se tocan.
    """
    con_lista = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type == 'list':
            for r in dv.sqref.ranges:
                for fila in ws[str(r)] if ':' in str(r) else [[ws[str(r)]]]:
                    for c in fila:
                        con_lista.add(c.coordinate)
    libres = set(getattr(ws, '_pf_negativos', set()))
    libres.update(NEGATIVOS.get(fname or '', {}).get(ws.title, []))
    importes, porcentajes, negativos = [], [], []
    for row in ws.iter_rows():
        for c in row:
            if not es_verde(c) or c.coordinate in con_lista:
                continue
            fmt = c.number_format or ''
            if c.coordinate in libres:
                negativos.append(c.coordinate)
            elif '%' in fmt:
                porcentajes.append(c.coordinate)
            elif '€' in fmt or fmt.startswith('#,##0') or fmt == '0':
                importes.append(c.coordinate)
    if importes:
        dv_numerica(ws, importes, minimo=0)
    if porcentajes:
        dv_numerica(ws, porcentajes, minimo=0, maximo=1)
    if negativos:
        # Sin mínimo: un saldo en descubierto es un dato, no un error.
        dv_numerica(ws, negativos, minimo=-1000000000000,
                    titulo='Importe no válido',
                    mensaje='Escribe un número (puede ser negativo).')
    if informe is not None and (importes or porcentajes or negativos):
        informe.append(ws.title + ': DV numérica en ' + str(len(importes))
                       + ' importes, ' + str(len(porcentajes))
                       + ' porcentajes y ' + str(len(negativos))
                       + ' celdas que admiten negativo')
    return len(importes), len(porcentajes)


# ==========================================================================
# Utilidades de rejilla (copiadas del motor de escandallos v2.0)
# ==========================================================================
# RT-03: la versión anterior era `(\$?)([A-Z]{1,3})(\$?)(\d+)` y NO distinguía
# una referencia local de una referencia a OTRA hoja: al insertar una fila,
# `Datos!C14` se convertía en `Datos!C15`. La corrupción era silenciosa —la
# fórmula seguía siendo sintácticamente válida— y caía justo sobre las hojas
# que la SPEC manda reestructurar. Ahora el patrón captura el prefijo de hoja
# (con o sin comillas) y el tramo `:REF` de un rango, y `_traducir_formula`
# devuelve intacto todo lo que venga cualificado.
RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')
RX_REF_CTX = re.compile(
    r"(?<![A-Za-z0-9_])"
    r"((?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"(\$?[A-Z]{1,3}\$?\d+)"
    r"(:\$?[A-Z]{1,3}\$?\d+)?"
    r"(?![A-Za-z0-9_(])")
RX_UNA_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)$')

CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _mover_una(ref, idx, eje):
    m = RX_UNA_REF.match(ref)
    if not m:
        return ref
    d1, col, d2, fila = m.groups()
    ci, fi = column_index_from_string(col), int(fila)
    if eje == 'col' and ci >= idx:
        col = get_column_letter(ci + 1)
    if eje == 'fila' and fi >= idx:
        fila = str(fi + 1)
    return d1 + col + d2 + fila


def _traducir_formula(valor, idx, eje, solo_hoja=None):
    """Desplaza las referencias de una fórmula al insertar fila/columna.

    `solo_hoja=None` → sólo las referencias SIN prefijo de hoja (las locales).
    `solo_hoja='Datos'` → sólo las que apuntan explícitamente a esa hoja: es lo
    que necesita `insertar_fila` para arreglar las fórmulas que OTRAS hojas
    hacen contra la hoja modificada (RT-04).
    """
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        hoja, ref, rango = m.group(1), m.group(2), m.group(3)
        nombre = None
        if hoja:
            nombre = hoja[:-1].strip("'")
        if solo_hoja is None:
            if hoja:
                return m.group(0)
        else:
            if nombre != solo_hoja:
                return m.group(0)
        fuera = (hoja or '') + _mover_una(ref, idx, eje)
        if rango:
            fuera += ':' + _mover_una(rango[1:], idx, eje)
        return fuera

    return RX_REF_CTX.sub(_sub, valor)


def _rangos_dv(ws):
    return [(dict((k, getattr(dv, k, None)) for k in CAMPOS_DV),
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**dict((k, v) for k, v in attrs.items()
                                   if v is not None))
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje) if idx else r)


def _desplazar_rango(ref, idx, eje):
    partes = ref.split(':')
    fuera = []
    for p in partes:
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        fuera.append(d1 + col + d2 + fila)
    return ':'.join(fuera)


def _propagar_a_otras_hojas(ws, idx, eje):
    """RT-04: reescribe las referencias que OTRAS hojas hacen a `ws`.

    `insertar_fila`/`insertar_columna` sólo traducían las fórmulas de la hoja
    que modificaban. Con el mapa nuevo del 03 (el SALDO FINAL baja de la fila
    25 a la 29) las 36 fórmulas de `Alertas` seguirían leyendo la fila 25 —y el
    fichero se vería perfectamente bien—. También se sueltan los gráficos: sus
    `Reference` son fijas y `graficos()` los reconstruye al final del pipeline
    (§1.8), así que dejarlos apuntando al mapa viejo sería peor que rehacerlos.
    """
    wb = ws.parent
    tocadas = 0
    if wb is None:
        return tocadas
    for otra in wb.worksheets:
        if otra is ws:
            continue
        for row in otra.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith('=')):
                    continue
                nueva = _traducir_formula(v, idx, eje, solo_hoja=ws.title)
                if nueva != v:
                    c.value = nueva
                    tocadas += 1
        if otra._charts:
            otra._charts = []
    ws._charts = []
    return tocadas


def insertar_columna(ws, idx):
    """Inserta una columna manteniendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos_ = dict((k, v.width) for k, v in ws.column_dimensions.items()
                   if v.width)

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra, ancho in sorted(anchos_.items(),
                               key=lambda kv: -column_index_from_string(kv[0])):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = ancho

    _propagar_a_otras_hojas(ws, idx, 'col')


def insertar_fila(ws, idx):
    """Equivalente por filas de `insertar_columna`."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = dict((k, v.height) for k, v in ws.row_dimensions.items()
                   if v.height)

    for fila in range(max_r, idx - 1, -1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila + 1, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'fila'))
    _restaurar_dv(ws, dvs, idx, 'fila')

    for fila, alto in sorted(alturas.items(), reverse=True):
        if fila >= idx:
            ws.row_dimensions[fila + 1].height = alto

    _propagar_a_otras_hojas(ws, idx, 'fila')


def anchos(ws, mapa):
    """`{'B': 45}` — anchos que no truncan (§1.6/TEC-29)."""
    for letra, ancho in mapa.items():
        actual = ws.column_dimensions[letra].width
        if actual is None or actual < ancho:
            ws.column_dimensions[letra].width = ancho


# ==========================================================================
# Impresión A4 (el censo exige paperSize 9 + fitToPage + pie)
# ==========================================================================
PIE = 'AI Chef Pro · aichef.pro · Página &P de &N'


def print_setup(ws, header_row=None, landscape=None, congelar=None):
    """A4 completo. CONSERVADOR a propósito: no cambia la orientación ni el
    `freeze_panes` que ya traía la hoja salvo que se le pidan explícitamente
    (los 10 ficheros ya venían con orientación y paneles bien puestos de la
    Fase A; reescribirlos sólo generaría diferencias sin motivo)."""
    ws.page_setup.paperSize = 9
    if landscape is not None:
        ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = PIE
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = str(header_row) + ':' + str(header_row)
    if congelar:
        ws.freeze_panes = congelar


# ==========================================================================
# Instrucciones (§1.7)
# ==========================================================================
def _col_texto(ws):
    """Columna donde vive el texto de `Instrucciones` (B en los 10)."""
    return 2 if ws.cell(row=2, column=2).value else 1


def linea_instrucciones(ws, texto, rx=None):
    """Sustituye la línea que case con `rx` o la añade al final. Nunca duplica.

    Los grupos la usan para reescribir instrucciones concretas (01b!B9,
    04!B13…) y para añadir la línea «de dónde sacar este dato» del §5.
    """
    col = _col_texto(ws)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return r
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return r
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)
    return destino


def cierre_instrucciones(ws, fname):
    """Bloque de cierre: nota de IVA, «Revisar → Desproteger», BIO y VERSIÓN.

    La bio es una **INSERCIÓN**: no la lleva ninguno de los 10 ficheros. Para
    que la 2.ª pasada no la vaya empujando hacia abajo, el bloque se BORRA
    entero primero y se reescribe al final del texto vivo — así su posición
    depende del contenido, no de dónde estaba el bloque en la pasada anterior,
    aunque un grupo haya añadido líneas nuevas por el medio.
    """
    col = _col_texto(ws)
    patrones = (RX_NOTA_IVA, RX_DESPROTEGER, RX_BIO, RX_VERSION)
    estilo = None
    for r in range(1, ws.max_row + 1):
        cel = ws.cell(row=r, column=col)
        v = cel.value
        if isinstance(v, str) and any(p.search(v) for p in patrones):
            if estilo is None:
                estilo = copy.copy(cel._style)
            cel.value = None
    ultima = 0
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value is not None:
            ultima = r
    lineas = []
    if fname not in SIN_NOTA_IVA:
        # RC-28: la checklist pre-apertura no tiene UNA sola cifra monetaria;
        # «Todas las cifras van SIN IVA» ahí es ruido que delata que la línea
        # se inyectó en bloque. El 04 tampoco la lleva: desde §3 desglosa base,
        # IVA y total con IVA en cada partida y lo dice en su cabecera.
        lineas.append(NOTA_IVA_CAJA if fname == FICHERO_CAJA else NOTA_IVA)
    lineas += [NOTA_DESPROTEGER, BIO_LINE, VERSION_LINE]
    fila = ultima + 2
    for texto in lineas:
        cel = ws.cell(row=fila, column=col, value=texto)
        if estilo is not None:
            cel._style = copy.copy(estilo)
        cel.alignment = Alignment(horizontal='left', vertical='top')
        fila += 1
    return fila - 1


# §1.7 — instrucciones que hoy son falsas o nombran pestañas inexistentes.
CORRECCIONES_INSTRUCCIONES = {
    '01b-plan-financiero-previsional-5-anos.xlsx': [
        # COM-28: «Año 1 con desglose mensual. Años 2-5 con desglose mensual.»
        # decía dos veces lo mismo y dejaba la frase sin sentido.
        (re.compile(r'^▸ Año 1 con desglose mensual'),
         '▸ Los cinco años van con desglose mensual (Ene-Dic) y total anual '
         'en la columna N.'),
    ],
    'BONUS-08-simulador-escenarios.xlsx': [
        # RC-12/COM-26: la variable del simulador se llama «Cubiertos / día»;
        # «Ocupación» no existe en ninguna celda. Es la superficie que más
        # manda, porque el comprador la lee con el fichero abierto.
        (re.compile(r'^▸ Cambia los valores de Ticket Medio'),
         '▸ Cambia los valores de Ticket medio, Cubiertos / día y Food Cost '
         'en la zona de inputs.'),
    ],
    '04-presupuesto-inversion-capex.xlsx': [
        # TEC-28: nombraba «Obra civil» y «Licencias y Permisos», pestañas que
        # no existen — las reales son «Obra» y «Licencias».
        (re.compile(r'^▸ Obra civil'),
         '▸ Obra, Equipamiento Cocina, Mobiliario Sala, Tecnología, '
         'Licencias.'),
    ],
}


def instrucciones(wb, fname, informe):
    if 'Instrucciones' not in wb.sheetnames:
        return
    ws = wb['Instrucciones']
    for rx, texto in CORRECCIONES_INSTRUCCIONES.get(fname, []):
        col = _col_texto(ws)
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and (rx.match(v) or v == texto):
                if v != texto:
                    ws.cell(row=r, column=col).value = texto
                    informe.append('Instrucciones!B' + str(r)
                                   + ': línea corregida (§1.7)')
                break
    anchos(ws, {'B': 100})


# ==========================================================================
# §1.5 — base «sin IVA» declarada en las etiquetas
# ==========================================================================
RX_VENTAS = re.compile(
    r'^(Ventas|Facturación|Ticket medio|Total Ingresos|TOTAL INGRESOS|'
    r'Comedor|Barra|Delivery|Eventos)', re.I)
SIN_IVA = ' (sin IVA)'

# §1.5 nombra exactamente estos ficheros: 01, 01b, 02, 05, 06 y 07. Se añade el
# BONUS-08 porque copia el ticket medio y la facturación del 02 —dejarlo sin
# marcar reproduce en él la incoherencia que el §1.5 viene a cerrar—, y se deja
# FUERA el 04: es un CAPEX, sus importes se desglosan con las columnas Base /
# IVA / Total del §3, y ahí «Barra» es un mueble de sala, no una línea de
# ingreso (la primera versión de esta función le puso «(sin IVA)» a
# `04!'Mobiliario Sala'!A7`, que es una barra de bar).
FICHEROS_SIN_IVA = (
    '01-plan-financiero-previsional.xlsx',
    '01b-plan-financiero-previsional-5-anos.xlsx',
    '02-calculadora-punto-equilibrio.xlsx',
    '05-pyl-mensual-real-vs-presupuesto.xlsx',
    '06-dashboard-ratios-financieros.xlsx',
    '07-informe-viabilidad-bancos.xlsx',
    'BONUS-08-simulador-escenarios.xlsx',
)


def _con_sin_iva(etiqueta):
    """«Ticket medio (€)» → «Ticket medio (€, sin IVA)», no «(€) (sin IVA)»."""
    if etiqueta.endswith(')'):
        return etiqueta[:-1] + ', sin IVA)'
    return etiqueta + SIN_IVA


def sin_iva(wb, fname, informe):
    """«(sin IVA)» en las etiquetas de ventas y de ticket (§1.5/DOM-17).

    El error número uno del sector: con el 10 % de restauración, un food cost
    calculado sobre ventas CON IVA sale ~3 puntos por debajo del real. El 03 ya
    lo aclaraba («IVA incl.»), así que el kit se contradecía consigo mismo.

    Se ejecuta en `cerrar()`, DESPUÉS de los grupos: así los grupos buscan sus
    etiquetas tal y como estaban («Facturación», «Ticket medio») sin tener que
    saber nada de este sufijo, y las filas que ELLOS creen también lo reciben.
    """
    if fname not in FICHEROS_SIN_IVA:
        return 0
    tocadas = 0
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            continue
        for col in (1, 2):
            for r in range(1, ws.max_row + 1):
                cel = ws.cell(row=r, column=col)
                v = cel.value
                if not isinstance(v, str) or not RX_VENTAS.match(v):
                    continue
                if 'IVA' in v:
                    continue
                # sólo etiquetas de fila/rótulo, nunca cabeceras de tabla ni
                # títulos: una cabecera lleva relleno oscuro.
                if cel.fill is not None and cel.fill.fill_type == 'solid':
                    rgb = cel.fill.fgColor.rgb
                    if isinstance(rgb, str) and rgb.upper().endswith(CAB):
                        continue
                cel.value = _con_sin_iva(v)
                tocadas += 1
    if tocadas:
        informe.append(fname + ': ' + str(tocadas)
                       + ' etiquetas de ventas/ticket marcadas «(sin IVA)»')
    return tocadas


# ==========================================================================
# §1.3 — guardas IFERROR en las divisiones desnudas
# ==========================================================================
# Celda → texto de ayuda que sustituye al `#¡DIV/0!`. Son las seis del 06 y las
# tres del 02 que la SPEC cita una por una: `02!Break-Even!C8` revienta con un
# coste variable del 100 %, `C9` con 0 días de apertura y `C10` con ticket 0.
GUARDAS = {
    # El 06 lo reescribe entero grupo_c con un mapa de filas nuevo y ya escribe
    # sus propias guardas; esto queda como red de seguridad para el caso
    # `--solo motor`, en el que la hoja sigue siendo la de la v1.1.
    '06-dashboard-ratios-financieros.xlsx': {
        'Ratios': {
            'C17': 'Indica las ventas',
            # §1.3 enumera SEIS divisiones desnudas en esta hoja (C17, C19,
            # C21, C22, C23, C24), pero son OCHO: `C18` (labor cost) y `C20`
            # (GOP) dividen por `C6` exactamente igual y revientan con ventas
            # a 0. La convención de familia manda —«IFERROR en toda
            # división»— así que van también.
            'C18': 'Indica las ventas',
            'C19': 'Indica las ventas',
            'C20': 'Indica las ventas',
            'C21': 'Indica los cubiertos',
            'C22': 'Indica plazas y horas',
            'C23': 'Indica los cubiertos',
            'C24': 'Indica las ventas',
        },
    },
    '02-calculadora-punto-equilibrio.xlsx': {
        'Break-Even': {
            'C8': 'Revisa el % de coste variable (no puede ser 100 %)',
            'C9': 'Indica los días de apertura',
            'C10': 'Indica el ticket medio',
        },
    },
}


def guardas(wb, fname, informe):
    """Envuelve en IFERROR las divisiones que hoy van desnudas (§1.3)."""
    puestas = 0
    for hoja, celdas in GUARDAS.get(fname, {}).items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        for coord, ayuda in celdas.items():
            v = ws[coord].value
            if not (isinstance(v, str) and v.startswith('=')):
                continue
            if 'IFERROR' in v.upper():
                continue
            nueva = iferror(v, ayuda)
            ws[coord].value = nueva
            reg(ws, coord, nueva)
            puestas += 1
    if puestas:
        informe.append(fname + ': ' + str(puestas)
                       + ' divisiones con IFERROR (§1.3)')
    return puestas


# ==========================================================================
# §1.4 — datos de ejemplo etiquetados
# ==========================================================================
# Hoja → fila donde va el rótulo del ejemplo (una fila POR ENCIMA del bloque
# precargado, que en los tres casos está vacía).
EJEMPLOS = {
    '02-calculadora-punto-equilibrio.xlsx': [('Datos', 3, 2)],
    '06-dashboard-ratios-financieros.xlsx': [('Ratios', 3, 2)],
    'BONUS-08-simulador-escenarios.xlsx': [('Simulador', 3, 1)],
    '03-cash-flow-forecast.xlsx': [('Flujo Mensual', 3, 1)],
}


def fila_ejemplo(ws, fila, col=1, texto=EJEMPLO_LINEA):
    """Rótulo «VALORES DE EJEMPLO — sustitúyelos por los tuyos» (§1.4).

    Un ejemplo sin marca acaba presentado al banco: es el mismo fallo que
    dejaba 150.000 € de inversión ficticia en el 07. Se escribe en ámbar sobre
    la fila vacía que hay encima del bloque precargado; si esa fila tuviera
    contenido, no se escribe nada (mejor sin rótulo que pisando un dato).
    """
    cel = ws.cell(row=fila, column=col)
    if cel.value not in (None, texto) and not RX_EJEMPLO.match(str(cel.value)):
        return False
    cel.value = texto
    cel.font = Font(bold=True, color=CF_AMBAR_FG, size=9)
    cel.alignment = Alignment(horizontal='left')
    return True


def ejemplos(wb, fname, informe):
    puestos = 0
    for hoja, fila, col in EJEMPLOS.get(fname, []):
        if hoja in wb.sheetnames and fila_ejemplo(wb[hoja], fila, col):
            puestos += 1
    if puestos:
        informe.append(fname + ': ' + str(puestos)
                       + ' hojas con rótulo de datos de ejemplo (§1.4)')
    return puestos


# ==========================================================================
# §1.6 — formatos, anchos y cabeceras
# ==========================================================================
ANCHOS = {
    '07-informe-viabilidad-bancos.xlsx': {
        # TEC-29: «Aval SGR (Sociedad de Garantía Recíproca)» se truncaba.
        'Garantías': {'B': 45, 'D': 40},
        'Resumen Ejecutivo': {'B': 34, 'C': 26},
        'Ratios': {'B': 34, 'C': 16, 'D': 24, 'E': 16},
    },
    '03-cash-flow-forecast.xlsx': {
        'Alertas': {'B': 26, 'C': 16, 'D': 22, 'E': 16},
    },
    '06-dashboard-ratios-financieros.xlsx': {
        'Ratios': {'B': 34, 'C': 16, 'D': 24, 'E': 16},
        'Benchmarks': {'B': 30},
    },
    '02-calculadora-punto-equilibrio.xlsx': {
        'Datos': {'B': 42, 'D': 46},
        'Break-Even': {'B': 40},
    },
}

# §1.6/TEC-20: `07!'Resumen Ejecutivo'` es la ÚNICA hoja del kit sin celdas
# creadas en su columna de valores y sin verde — el analista del banco lee
# «150000» donde debería leer «150.000,00 €», y no sabe dónde escribir.
RESUMEN_EJECUTIVO_FMT = {
    'C8': FMT_ENT, 'C9': FMT_ENT, 'C10': FMT_FECHA,
    'C13': FMT_EUR, 'C14': FMT_EUR, 'C15': FMT_EUR,
    'C16': '0.0" años"', 'C17': FMT_EUR, 'C18': FMT_EUR,
    'C19': '0.0" años"',
}


def formatos(wb, fname, informe):
    """Anchos + creación de la columna de valores del 07 (§1.6)."""
    for hoja, mapa in ANCHOS.get(fname, {}).items():
        if hoja in wb.sheetnames:
            anchos(wb[hoja], mapa)
    if fname != '07-informe-viabilidad-bancos.xlsx':
        return
    if 'Resumen Ejecutivo' not in wb.sheetnames:
        return
    ws = wb['Resumen Ejecutivo']
    creadas = 0
    for r in range(5, 26):
        etiqueta = ws.cell(row=r, column=2).value
        if not isinstance(etiqueta, str) or not etiqueta:
            continue
        if etiqueta.isupper():          # los tres epígrafes de bloque
            continue
        cel = ws.cell(row=r, column=3)
        cel.number_format = RESUMEN_EJECUTIVO_FMT.get(cel.coordinate, 'General')
        if not es_verde(cel):
            cel.fill = PatternFill('solid', fgColor=VERDE)
            creadas += 1
        cel.protection = Protection(locked=False)
        cel.alignment = Alignment(horizontal='left')
    ws.freeze_panes = 'C5'
    if creadas:
        informe.append("07!'Resumen Ejecutivo'!C5:C25: " + str(creadas)
                       + ' celdas creadas en verde con formato (§1.6)')


# ==========================================================================
# §1.1 — gráficos
# ==========================================================================
def _hay_datos(ws, coord):
    """La celda existe y tiene algo (valor o fórmula)."""
    try:
        return ws[coord].value is not None
    except Exception:                                        # noqa: BLE001
        return False


def _barras(ws, titulo, y_titulo, datos, categorias, ancla, from_rows=True):
    ch = BarChart()
    ch.type = 'col'
    ch.style = 10
    ch.title = titulo
    ch.y_axis.title = y_titulo
    ch.add_data(datos, titles_from_data=True, from_rows=from_rows)
    ch.set_categories(categorias)
    ch.width, ch.height = 18, 9
    ws.add_chart(ch, ancla)
    return ch


def _lineas(ws, titulo, y_titulo, datos, categorias, ancla, from_rows=True):
    ch = LineChart()
    ch.style = 12
    ch.title = titulo
    ch.y_axis.title = y_titulo
    ch.add_data(datos, titles_from_data=True, from_rows=from_rows)
    ch.set_categories(categorias)
    ch.width, ch.height = 18, 9
    ws.add_chart(ch, ancla)
    return ch


def _es_formula(ws, coord):
    try:
        v = ws[coord].value
    except Exception:                                        # noqa: BLE001
        return False
    return isinstance(v, str) and v.startswith('=')


def graficos(wb, fname, informe):
    """§1.1 — «gráficos automáticos» se anuncia SEIS veces en la landing y en
    la v1.1 había 0 charts en las 47 hojas. Aquí se construyen de verdad en 9
    de los 10 (el BONUS-09 es una checklist: no tiene serie que dibujar).

    Tres reglas que la v2.0 aprendió por las malas:

    1. **Precondición.** Si el bloque de datos que el gráfico necesita todavía
       no existe se OMITE y se anota como pendiente. Un gráfico sobre celdas
       vacías es peor que ninguno: el cliente lo ve plano en cero y cree que su
       negocio lo está (RC-10: los tres «Resumen» seguían a 0 y el motor les
       colgó encima un BarChart).
    2. **Ninguna serie de TEXTO** (RT-11/RD-26). El LineChart del 03 se
       construía sobre `C:E` de un tirón e incluía la columna `D`, que es
       «OK» / «⚠️ BAJO MÍNIMO»: Excel la dibuja como una línea plana en cero
       con leyenda «Alerta», dentro del único gráfico que el cliente enseña.
       Las series no contiguas se añaden con dos `Reference`.
    3. **Anclados DEBAJO de la tabla** (RT-12). Con `fitToWidth=1`, un gráfico
       a la derecha obliga a Excel a escalar TODA la hoja para que tabla +
       gráfico entren en el ancho de un A4: medido, el 03 y el 04 se imprimían
       al 48-49 % y el 07 al 62 %. En un kit cuyo listón es «que el 07 se
       imprima y se entregue», eso empeoraba la impresión.
    """
    hechos, pendientes = [], []

    def _limpiar(ws):
        ws._charts = []

    if fname in ('01-plan-financiero-previsional.xlsx',
                 '01b-plan-financiero-previsional-5-anos.xlsx'):
        ws = wb['Resumen'] if 'Resumen' in wb.sheetnames else None
        ultima = 4 if fname.startswith('01-') else 6      # col D o col F
        if ws is None:
            pendientes.append('Resumen: la hoja no existe')
        elif not _es_formula(ws, 'B5'):
            pendientes.append('Resumen!B5 no consolida todavía: es de grupo_a')
        else:
            _limpiar(ws)
            datos = Reference(ws, min_col=1, max_col=ultima, min_row=5,
                              max_row=7)
            cats = Reference(ws, min_col=2, max_col=ultima, min_row=4)
            _barras(ws, 'Ingresos, gastos y EBITDA por año', '€', datos, cats,
                    'A11')
            hechos.append('Resumen!A11 BarChart ingresos/gastos/EBITDA')

    elif fname == '02-calculadora-punto-equilibrio.xlsx':
        ws = wb['Break-Even'] if 'Break-Even' in wb.sheetnames else None
        if ws is None or not _hay_datos(ws, 'A20'):
            pendientes.append('Break-Even!A20:D28: el bloque auxiliar de '
                              'ingresos vs costes lo crea grupo_b')
        else:
            _limpiar(ws)
            datos = Reference(ws, min_col=2, max_col=4, min_row=20, max_row=28)
            cats = Reference(ws, min_col=1, min_row=21, max_row=28)
            _lineas(ws, 'Ingresos vs costes totales por cubiertos/día', '€',
                    datos, cats, 'A31', from_rows=False)
            hechos.append('Break-Even!A31 LineChart ingresos vs costes')

    elif fname == FICHERO_CAJA:
        ws = wb['Alertas'] if 'Alertas' in wb.sheetnames else None
        if ws is None or not _hay_datos(ws, 'C6') or not _hay_datos(ws, 'E6'):
            pendientes.append('Alertas!C6:C17 o la serie de umbral E6:E17')
        else:
            _limpiar(ws)
            ch = LineChart()
            ch.style = 12
            ch.title = 'Saldo final vs umbral de seguridad'
            ch.y_axis.title = '€'
            # SÓLO C (saldo) y E (umbral): la D es texto (RT-11/RD-26).
            ch.add_data(Reference(ws, min_col=3, max_col=3, min_row=5,
                                  max_row=17), titles_from_data=True)
            ch.add_data(Reference(ws, min_col=5, max_col=5, min_row=5,
                                  max_row=17), titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=2, min_row=6, max_row=17))
            ch.width, ch.height = 18, 9
            ws.add_chart(ch, 'B20')
            hechos.append('Alertas!B20 LineChart saldo + umbral (sin la '
                          'columna de texto)')

    elif fname == '04-presupuesto-inversion-capex.xlsx':
        ws = wb['Resumen'] if 'Resumen' in wb.sheetnames else None
        if ws is None or not _es_formula(ws, 'E5'):
            pendientes.append('Resumen!E5 (columna Real del mapa nuevo): la '
                              'crea grupo_b')
        else:
            _limpiar(ws)
            ch = BarChart()
            ch.type = 'col'
            ch.style = 10
            ch.title = 'Presupuesto (base) vs real por categoría'
            ch.y_axis.title = '€'
            ch.add_data(Reference(ws, min_col=2, max_col=2, min_row=4,
                                  max_row=9), titles_from_data=True)
            ch.add_data(Reference(ws, min_col=5, max_col=5, min_row=4,
                                  max_row=9), titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=9))
            ch.width, ch.height = 18, 9
            ws.add_chart(ch, 'A18')
            hechos.append('Resumen!A18 BarChart presupuesto vs real')

    elif fname == '05-pyl-mensual-real-vs-presupuesto.xlsx':
        ws = wb['Resumen Anual'] if 'Resumen Anual' in wb.sheetnames else None
        if ws is None:
            pendientes.append("'Resumen Anual' no existe")
        elif not _es_formula(ws, 'B5'):
            pendientes.append("'Resumen Anual'!B5 no consolida todavía: es de "
                              'grupo_a')
        else:
            _limpiar(ws)
            datos = Reference(ws, min_col=1, max_col=13, min_row=5, max_row=6)
            cats = Reference(ws, min_col=2, max_col=13, min_row=4)
            _barras(ws, 'Ingresos previstos vs reales por mes', '€', datos,
                    cats, 'A27')
            hechos.append("'Resumen Anual'!A27 BarChart previsto vs real")

    elif fname == '06-dashboard-ratios-financieros.xlsx':
        ws = wb['Ratios'] if 'Ratios' in wb.sheetnames else None
        bm = wb['Benchmarks'] if 'Benchmarks' in wb.sheetnames else None
        numerico = bm is not None and isinstance(
            bm['F4'].value, (int, float))
        if ws is None or not numerico or not _es_formula(ws, 'F17'):
            pendientes.append('Benchmarks!F:G (columnas numéricas «óptimo»/'
                              '«límite») y Ratios!F17: las crea grupo_c')
        else:
            _limpiar(ws)
            ch = BarChart()
            ch.type = 'col'
            ch.style = 10
            ch.title = 'Tu valor vs objetivo del sector'
            ch.y_axis.title = '%'
            ch.add_data(Reference(ws, min_col=3, max_col=3, min_row=16,
                                  max_row=20), titles_from_data=True)
            ch.add_data(Reference(ws, min_col=6, max_col=6, min_row=16,
                                  max_row=20), titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=2, min_row=17, max_row=20))
            ch.width, ch.height = 18, 9
            ws.add_chart(ch, 'A42')
            hechos.append('Ratios!A42 BarChart tu valor vs objetivo')

    elif fname == '07-informe-viabilidad-bancos.xlsx':
        ws = wb['Proyecciones'] if 'Proyecciones' in wb.sheetnames else None
        if ws is None:
            pendientes.append('Proyecciones no existe')
        elif not _es_formula(ws, 'B11') or not _es_formula(ws, 'B18'):
            pendientes.append('Proyecciones!B11 (EBITDA) y B18 (cash flow) '
                              'siguen siendo constantes: son de grupo_c')
        else:
            _limpiar(ws)
            ch = LineChart()
            ch.style = 12
            ch.title = 'EBITDA y cash flow operativo (años 1-5)'
            ch.y_axis.title = '€'
            ch.add_data(Reference(ws, min_col=1, max_col=6, min_row=11,
                                  max_row=11), titles_from_data=True,
                        from_rows=True)
            ch.add_data(Reference(ws, min_col=1, max_col=6, min_row=18,
                                  max_row=18), titles_from_data=True,
                        from_rows=True)
            ch.set_categories(Reference(ws, min_col=2, max_col=6, min_row=3))
            ch.width, ch.height = 18, 9
            ws.add_chart(ch, 'A34')
            hechos.append('Proyecciones!A34 LineChart EBITDA + cash flow')

    elif fname == 'BONUS-08-simulador-escenarios.xlsx':
        ws = wb['Comparativa'] if 'Comparativa' in wb.sheetnames else None
        if ws is None:
            pendientes.append('Comparativa no existe')
        else:
            _limpiar(ws)
            datos = Reference(ws, min_col=2, max_col=5, min_row=7, max_row=7)
            cats = Reference(ws, min_col=3, max_col=5, min_row=3)
            _barras(ws, 'EBITDA anual por escenario', '€', datos, cats, 'B11')
            hechos.append('Comparativa!B11 BarChart EBITDA por escenario')

    for h in hechos:
        informe.append(fname + ': gráfico ' + h)
    for p in pendientes:
        informe.append(fname + ': gráfico PENDIENTE — ' + p)
    return {'hechos': hechos, 'pendientes': pendientes}


# ==========================================================================
# §1.2 — semáforos que el motor puede poner ya (los rangos que hoy existen)
# ==========================================================================
def semaforos(wb, fname, informe):
    """Formato condicional sobre los rangos de estado.

    RT-17: cada regla se condiciona a que su rango tenga al menos una celda con
    contenido. Antes se escribían igual sobre rangos SIN una sola celda
    —`07!Ratios!E4:E10` en una hoja que terminaba en la columna D— y el gate
    contaba reglas, no efectos: 3 reglas que no podían encenderse jamás
    figuraban como trabajo hecho.
    """
    puestos, pendientes = [], []

    def _poner(ws, rango, vocab, etiqueta):
        if semaforo(ws, rango, vocab):
            puestos.append(etiqueta)
            return True
        pendientes.append(etiqueta + ' (rango sin celdas: no se escribe)')
        return False

    if fname == FICHERO_CAJA and 'Alertas' in wb.sheetnames:
        ws = wb['Alertas']
        _poner(ws, 'D6:D17', SEM_TESORERIA, 'Alertas!D6:D17')
        # §1.2: además de la palabra, el saldo bajo umbral se pinta él mismo.
        # El `<>"—"` evita que se pinte de rojo un mes SIN DATOS (RD-25): el
        # fichero salía con once meses en «BAJO MÍNIMO» ya cacheados.
        regla_expresion(ws, 'C6:C17', '=AND($C6<$C$3,$D6<>"—")')
        # Serie de umbral del gráfico (§1.1): columna E = el umbral repetido.
        for r in range(6, 18):
            f(ws, 'E' + str(r), '=$C$3', FMT_EUR)
        ws['E5'] = 'Umbral'
        ws['E5']._style = copy.copy(ws['D5']._style)
        puestos.append('Alertas!C6:C17 (expresión) + serie de umbral E6:E17')

    if fname == '05-pyl-mensual-real-vs-presupuesto.xlsx':
        for mes in ('Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago',
                    'Sep', 'Oct', 'Nov', 'Dic'):
            if mes in wb.sheetnames:
                _poner(wb[mes], 'F6:F30', SEM_DESVIACION, mes + '!F6:F30')

    if fname == '06-dashboard-ratios-financieros.xlsx' \
            and 'Ratios' in wb.sheetnames:
        _poner(wb['Ratios'], 'E17:E25', SEM_RATIO + SEM_GOP, 'Ratios!E17:E25')

    if fname == '07-informe-viabilidad-bancos.xlsx' \
            and 'Ratios' in wb.sheetnames:
        ws = wb['Ratios']
        _poner(ws, 'E4:E10', SEM_BANCO, 'Ratios!E4:E10')
        # RT-18: §1.2 pedía regla de EXPRESIÓN «para el DSCR», que es el número
        # que decide la operación bancaria, y sólo había tres containsText.
        if isinstance(ws['G5'].value, (int, float)):
            regla_expresion(ws, 'C5:C5', '=$C$5<$G$5')
            puestos.append('Ratios!C5 (expresión contra el 1,25x del DSCR)')
        else:
            pendientes.append('Ratios!C5: el umbral numérico G5 lo crea '
                              'grupo_c')

    if fname == 'BONUS-09-checklist-pre-apertura.xlsx' \
            and 'Checklist' in wb.sheetnames:
        _poner(wb['Checklist'], 'F5:F64', SEM_ESTADO, 'Checklist!F5:F64')

    for p in puestos:
        informe.append(fname + ': formato condicional ' + p)
    for p in pendientes:
        informe.append(fname + ': formato condicional PENDIENTE — ' + p)
    return puestos


# ==========================================================================
# Metadata (§1.7)
# ==========================================================================
def metadatos(wb, fname, titulo=None):
    p = wb.properties
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'
    p.subject = SUBJECT_V2
    p.keywords = KEYWORDS_V2
    if titulo:
        p.title = titulo
    elif isinstance(p.title, str):
        p.title = re.sub(r'· v\d+\.\d+', '· v2.0', p.title)
    return p.title


# ==========================================================================
# Protección (§1.3)
# ==========================================================================
def proteger(ws, informe=None):
    """Protección SIN contraseña: se desbloquean SOLO las celdas verdes.

    Ojo con `password`: `= None` revienta openpyxl y `= ''` escribe el hash de
    la cadena vacía → Excel pide contraseña justo donde las Instrucciones dicen
    que no hay ninguna. Se deja sin asignar.
    """
    extras = getattr(ws, '_pf_editables', set())
    verdes = 0
    for row in ws.iter_rows():
        for c in row:
            if es_verde(c) or c.coordinate in extras:
                c.protection = Protection(locked=False)
                verdes += 1
            else:
                c.protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.sort = False
    if informe is not None:
        informe.append(ws.title + ': protegida sin contraseña (' + str(verdes)
                       + ' celdas verdes editables)')
    return verdes


# ==========================================================================
# Auditoría auxiliar: parámetros literales dentro de fórmulas (§1.3)
# ==========================================================================
RX_LITERAL = re.compile(r'(?<![A-Z0-9_$.!])(\d+(?:\.\d+)?)(?![0-9.]*[%)]?\s*'
                        r'[A-Z]*\d*\()')
LITERALES_TOLERADOS = {'0', '1', '2', '3', '4', '5', '12', '30', '100',
                       '0.0', '1.0'}


def literales_sospechosos(wb, fname):
    """Fórmulas con un número clavado dentro que TAMBIÉN existe como input.

    «Ningún parámetro que exista como input en otra hoja del libro va como
    literal en una fórmula» (§1.3). Esto no corrige: informa, para que el gate
    del orquestador vea de un vistazo si queda alguno (el `26` de
    `02!Escenarios!C10` y el `0.08` del VAN del 07 son los casos citados).
    """
    fuera = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith('=')):
                    continue
                cuerpo = re.sub(r'"[^"]*"', '', v)
                for m in RX_LITERAL.finditer(cuerpo):
                    num = m.group(1)
                    if num in LITERALES_TOLERADOS:
                        continue
                    fuera.append({'fichero': fname, 'hoja': ws.title,
                                  'celda': c.coordinate, 'literal': num,
                                  'formula': v[:90]})
                    break
    return fuera


RX_GUARDA = re.compile(r'^=\s*(IFERROR|IF)\s*\(', re.I)


def contadores(wb, fname):
    """Contador auxiliar del estado de un libro (alimenta el informe/gates)."""
    r = {'fichero': fname, 'hojas': len(wb.worksheets), 'formulas': 0,
         'verdes': 0, 'charts': 0, 'cf': 0, 'dv': 0, 'protegidas': 0,
         'ceros_constantes': 0, 'divisiones_sin_iferror': 0, 'vacias': 0}
    for ws in wb.worksheets:
        r['charts'] += len(ws._charts)
        r['cf'] += sum(len(cf.rules) for cf in ws.conditional_formatting)
        r['dv'] += len(ws.data_validations.dataValidation)
        if ws.protection.sheet:
            r['protegidas'] += 1
        for row in ws.iter_rows():
            for c in row:
                if es_verde(c):
                    r['verdes'] += 1
                v = c.value
                if isinstance(v, str) and v.startswith('='):
                    r['formulas'] += 1
                    if '/' in v and not RX_GUARDA.match(v):
                        # RT-14: la guarda anterior buscaba la subcadena
                        # «IF(», que aparece dentro de COUNTIF( y SUMIF(. Toda
                        # división escrita con COUNTIF —que la SPEC IMPONE
                        # porque pycel no implementa COUNTA— se contaba como
                        # protegida sin estarlo: BONUS-09!C59 daba 0.
                        r['divisiones_sin_iferror'] += 1
                elif v == 0 and not es_verde(c):
                    # 0 constante SIN relleno de input = resultado en cero, el
                    # patrón exacto de los tres «Resumen» que no consolidan.
                    r['ceros_constantes'] += 1
                elif v == '':
                    r['vacias'] += 1
    return r


# ==========================================================================
# §1.8 — TIR por Newton-Raphson y caché en el XML
# ==========================================================================
def van(tasa, flujos):
    """VAN(r) = Σ Fᵢ/(1+r)ⁱ con el flujo del año 0 en `flujos[0]`."""
    total = 0.0
    for i, fl in enumerate(flujos):
        total += float(fl) / ((1.0 + tasa) ** i)
    return total


def _dvan(tasa, flujos):
    total = 0.0
    for i, fl in enumerate(flujos):
        if i:
            total += -i * float(fl) / ((1.0 + tasa) ** (i + 1))
    return total


def _biseccion(flujos, lo=-0.99, hi=10.0, iteraciones=200, tol=1e-10):
    flo, fhi = van(lo, flujos), van(hi, flujos)
    if flo * fhi > 0:
        return None
    for _ in range(iteraciones):
        medio = (lo + hi) / 2.0
        fm = van(medio, flujos)
        if abs(fm) < tol or (hi - lo) < tol:
            return medio
        if flo * fm < 0:
            hi, fhi = medio, fm
        else:
            lo, flo = medio, fm
    return (lo + hi) / 2.0


def tir_newton(flujos, semilla=0.1, iteraciones=100, tol=1e-10):
    """TIR por Newton-Raphson; bisección en [−0,99, 10] si la derivada se anula.

    pycel **no implementa IRR** («Function IRR is not implemented. IRR is in
    the "Financial" group») y su `IFERROR` no lo atrapa: la evaluación revienta
    antes, así que `inject_cache.py` deja `07!Proyecciones!B21` sin `<v>` y en
    Vista previa de macOS o en Google Sheets la celda sale EN BLANCO — en el
    documento que se entrega al banco, justo en la línea que decide la
    operación. Caso trazado de la SPEC: −150.000 / 30.000 / 45.000 / 60.000 /
    70.000 → 11,9592 %.

    Devuelve `None` cuando no hay cambio de signo (todo ceros, o todo
    positivo): ahí la TIR no existe y lo honesto es dejar el «—» de la
    fórmula, no inventar un número.
    """
    limpios = [float(x or 0) for x in flujos]
    if not limpios or min(limpios) >= 0 or max(limpios) <= 0:
        return None
    r = semilla
    for _ in range(iteraciones):
        v = van(r, limpios)
        if abs(v) < tol:
            return r
        d = _dvan(r, limpios)
        if abs(d) < 1e-14:
            return _biseccion(limpios)
        nuevo = r - v / d
        if nuevo <= -0.9999:
            nuevo = (r - 0.9999) / 2.0
        if nuevo > 1e6:
            return _biseccion(limpios)
        r = nuevo
    return _biseccion(limpios)


def payback(flujos):
    """Payback en años con la parte fraccionaria del año en que cruza 0."""
    acum = 0.0
    previo = 0.0
    for i, fl in enumerate(flujos):
        previo = acum
        acum += float(fl or 0)
        if i and acum >= 0 and previo < 0:
            return (i - 1) + (-previo / float(flujos[i] or 1))
    return None


# --- inyección del valor en el XML (mismo mecanismo que inject_cache.py) ---
def _mapa_hojas(z):
    wbxml = z.read('xl/workbook.xml').decode('utf-8')
    rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    relmap = {}
    for rel in re.findall(r'<Relationship[^>]*/>', rels):
        rid = re.search(r'Id="([^"]+)"', rel)
        tgt = re.search(r'Target="([^"]+)"', rel)
        if rid and tgt:
            relmap[rid.group(1)] = tgt.group(1)
    fuera = {}
    for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"',
                         wbxml):
        tgt = relmap.get(m.group(2), '')
        if tgt:
            sfile = tgt.lstrip('/') if tgt.startswith('/') else 'xl/' + tgt
            fuera[html.unescape(m.group(1))] = sfile
    return fuera


def inyectar_valor(path, hoja, coord, valor):
    """Escribe el `<v>` de una celda de fórmula sin pasar por `wb.save()`.

    Copiado del mecanismo de `inject_cache.py` (reescritura del zip a mano):
    cualquier guardado posterior de openpyxl borraría el caché, así que esto
    va SIEMPRE al final del pipeline.
    """
    z = zipfile.ZipFile(path)
    hojas = _mapa_hojas(z)
    sfile = hojas.get(hoja)
    parts = dict((n, z.read(n)) for n in z.namelist())
    z.close()
    if sfile is None or sfile not in parts:
        return False
    xml = parts[sfile].decode('utf-8')
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        pat = re.compile(r'(<c r="' + coord + r'"[^>]*>)(<f[^>]*>[^<]*</f>)'
                         r'(?:<v>[^<]*</v>|<v\s*/>)?(</c>)')
        rep = r'\1\2<v>' + repr(float(valor)) + r'</v>\3'
        xml, n = pat.subn(rep, xml, count=1)
    else:
        esc = html.escape(str(valor))
        pat = re.compile(r'<c r="' + coord + r'"([^>]*)>(<f[^>]*>[^<]*</f>)'
                         r'(?:<v>[^<]*</v>|<v\s*/>)?</c>')

        def _rep(m):
            attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
            return ('<c r="' + coord + '"' + attrs + ' t="str">' + m.group(2)
                    + '<v>' + esc + '</v></c>')

        xml, n = pat.subn(_rep, xml, count=1)
    if not n:
        return False
    parts[sfile] = xml.encode('utf-8')
    tmp = path + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for nombre, data in parts.items():
            zout.writestr(nombre, data)
    shutil.move(tmp, path)
    return True


# §4 reordena el 07: el P&L crece con «Coste de bebida» y «Comisiones de
# plataformas» (RD-13/RD-14) y el bloque de flujos gana la columna del AÑO 0,
# que antes no existía —`NPV(0.08,C19:F19)+B19` sólo tomaba cuatro años de
# cinco—. De ahí las direcciones nuevas.
CELDA_IRR = ('Proyecciones', 'B24')
RANGO_FLUJOS = ('Proyecciones', 'B22', 'G22')
CELDA_TASA = 'C27'


def cachear_irr(path, informe=None):
    """Calcula la TIR con los flujos YA cacheados y la inyecta (§1.8).

    Se ejecuta **después** de `inject_cache.py`: lee `Proyecciones!B19:F19` con
    `data_only=True` (valores que pycel acaba de dejar) y escribe el `<v>` de
    `B21`. Si no hay cambio de signo, inyecta el texto «—» de la propia
    fórmula: mejor un guion honesto que un número inventado.
    """
    import openpyxl
    fuera = {'fichero': path.rsplit('/', 1)[-1], 'flujos': None, 'tir': None,
             'van': None, 'payback': None, 'inyectada': False, 'motivo': None}
    wb = openpyxl.load_workbook(path, data_only=True)
    hoja = RANGO_FLUJOS[0]
    if hoja not in wb.sheetnames:
        fuera['motivo'] = 'no existe la hoja ' + hoja
        return fuera
    ws = wb[hoja]
    flujos = []
    col0 = column_index_from_string(re.match(r'([A-Z]+)', RANGO_FLUJOS[1])
                                    .group(1))
    col1 = column_index_from_string(re.match(r'([A-Z]+)', RANGO_FLUJOS[2])
                                    .group(1))
    fila = int(re.search(r'(\d+)', RANGO_FLUJOS[1]).group(1))
    for c in range(col0, col1 + 1):
        v = ws.cell(row=fila, column=c).value
        flujos.append(float(v) if isinstance(v, (int, float)) else 0.0)
    fuera['flujos'] = flujos
    tasa = None
    hoja_irr, coord_irr = CELDA_IRR
    wbf = openpyxl.load_workbook(path)
    if hoja_irr in wbf.sheetnames:
        # la tasa de descuento del VAN vive en celda desde la v2.0 (C25)
        celda_tasa = wbf[hoja_irr][CELDA_TASA].value
        if isinstance(celda_tasa, (int, float)):
            tasa = float(celda_tasa)
    tir = tir_newton(flujos)
    fuera['tir'] = tir
    fuera['van'] = van(tasa if tasa is not None else 0.08, flujos)
    fuera['payback'] = payback(flujos)
    # RT-13: el literal alterno se LEE de la propia fórmula. Antes se inyectaba
    # un «—» fijo mientras la celda decía `IFERROR(IRR(...),"N/A")`: quien
    # abriera el fichero sin recalcular (Vista previa de macOS, Google Sheets,
    # un conversor a PDF) leía «—» y quien recalculara en Excel leía «N/A».
    alterna = '—'
    if hoja_irr in wbf.sheetnames:
        formula = wbf[hoja_irr][coord_irr].value
        if isinstance(formula, str):
            m = re.search(r',\s*"([^"]*)"\s*\)\s*$', formula)
            if m:
                alterna = m.group(1)
    fuera['alterna_de_la_formula'] = alterna
    valor = tir if tir is not None else alterna
    if tir is None:
        fuera['motivo'] = ('sin cambio de signo en los flujos: la TIR no '
                           'existe (el libro se entrega con el ejemplo a 0)')
    fuera['inyectada'] = inyectar_valor(path, hoja_irr, coord_irr, valor)
    if informe is not None:
        informe.append('TIR cacheada: ' + repr(valor) + ' en ' + hoja_irr
                       + '!' + coord_irr
                       + (' (OK)' if fuera['inyectada'] else ' (NO inyectada)'))
    return fuera


# ==========================================================================
# API principal: aplicar (§1 antes de los grupos) y cerrar (§1 después)
# ==========================================================================
def aplicar(wb, fname, informe):
    """§1 transversal ANTES del trabajo de los grupos."""
    metadatos(wb, fname)
    instrucciones(wb, fname, informe)
    guardas(wb, fname, informe)
    formatos(wb, fname, informe)
    ejemplos(wb, fname, informe)
    return informe


def cerrar(wb, fname, informe, proteger_hojas=True):
    """Cierre común: limpieza, semáforos, DV, impresión A4, protección y el
    bloque de bio + versión de `Instrucciones`.

    Va DESPUÉS de los grupos: si se protegiera antes, cada celda que un grupo
    creara después nacería bloqueada aunque fuese verde.
    """
    sin_iva(wb, fname, informe)
    semaforos(wb, fname, informe)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    # el censo cuenta `''` como defecto `empty_str`
                    c.value = None
        if ws.title == 'Instrucciones':
            print_setup(ws, None, landscape=False)
            cierre_instrucciones(ws, fname)
            if proteger_hojas:
                proteger(ws, informe)
            continue
        validaciones(ws, informe, fname)
        print_setup(ws)
        if proteger_hojas:
            proteger(ws, informe)
    metadatos(wb, fname)
    return informe
