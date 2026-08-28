#!/usr/bin/env python3
"""
grupo_a.py — §2 de la SPEC: proyección y consolidados (01, 01b, 05).

Qué arregla, id por id (ronda 1 + refutadores de la ronda 2):

  · DOM-04 / TEC-05 / RD-09 / RC-10 — `01!Resumen` y `01b!Resumen` eran 0
    CONSTANTES sin relleno de input mientras las Instrucciones prometían que
    «consolida los 3/5 años». Ahora consolidan **por referencia** a la columna
    TOTAL AÑO de cada pestaña, que es el patrón que el propio kit ya usaba en
    `04!Resumen`.
  · DOM-05 / TEC-04 / RD-08 — `05!'Resumen Anual'` eran 78 constantes 0 y la
    columna TOTAL tampoco sumaba; encima el motor colgaba un BarChart de esas
    constantes (gráfico permanentemente plano). Ahora las 12 pestañas se
    consolidan por referencia y `N` suma.
  · DOM-27 / RD-09 — la columna de crecimiento sólo tenía fórmula en `E5`
    (`G5` en el 01b): `E6:E7` replican el crecimiento y `E8` va en **p.p.**,
    porque un margen no «crece un %», varía en puntos.
  · DOM-12 / TEC-09 / TEC-10 / RD-27 / RT-10 / RC-18 — el semáforo del 05 usaba
    `ABS()`: vender un 20 % POR ENCIMA del presupuesto salía «🔴 Alerta». Ahora
    el criterio va **con signo** y por familia de fila (ingresos y EBITDA:
    quedarse corto es lo malo; gastos: pasarse es lo malo). Y `E` del EBITDA
    divide entre `ABS()` del presupuesto, porque con EBITDA presupuestado
    negativo —lo normal en una apertura— el signo se invertía y una mejora de
    3.000 € salía en rojo.
  · TEC-26 / RC-19 — media tabla vacía bajo su propio encabezado: «Margen
    EBITDA %» y los tres ratios de «RATIOS AUTOMÁTICOS» sólo tenían
    Presupuesto y Real. Ahora tienen Desviación (en **p.p.**), la marca «p.p.»
    y semáforo propio.
  · RD-13 (hallazgo nuevo del refutador de dominio) — el kit desglosaba el
    INGRESO de delivery en las cuatro plantillas de P&L y no tenía **ni una**
    línea de gasto por la comisión de las plataformas (30-35 % de esa venta).
    Un restaurante con un 20 % de venta en delivery se llevaba al banco un
    EBITDA inflado en 6-7 puntos. Línea nueva en 01, 01b y 05 (y en el 07 y el
    03, que los hacen grupo_c y grupo_b).
  · RD-14 (hallazgo nuevo) — el Food Cost % se calculaba sobre el TOTAL de
    ingresos, barra incluida, y se comparaba con un benchmark de food cost
    puro. Se desdobla el coste en **comida** y **bebida**: Food Cost % =
    coste de comida / (ventas − barra) y Beverage Cost % = coste de bebida /
    ventas de barra, cada uno contra su propio umbral.
  · RT-25 (parcial) — los umbrales del semáforo (5 % / 10 % y 2 p.p. / 5 p.p.)
    NO van clavados dentro de la fórmula: viven en celdas verdes de
    `'Resumen Anual'`, que es lo que exige §1.3 y lo que hace que el gate de
    literales pueda cerrarse en 0.

Convenciones: fórmulas con `motor.f` (quedan registradas y `main.py` verifica
una por una que acabaron con caché), editables con `motor.val(..., verde_=True)`
—el verde es lo que `motor.cerrar()` usa para decidir qué se desbloquea—, y
ninguna división desnuda. `motor.limpiar_rango` antes de reescribir un mapa de
filas: si el verde de la v1.1 sobrevive en una celda que ya no es input,
`proteger()` la deja abierta y `validaciones()` le cuelga una DV que no le toca.
"""
import copy

from openpyxl.styles import Alignment, Font

import motor

FICHEROS = [
    '01-plan-financiero-previsional.xlsx',
    '01b-plan-financiero-previsional-5-anos.xlsx',
    '05-pyl-mensual-real-vs-presupuesto.xlsx',
]

MESES = ('Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct',
         'Nov', 'Dic')

# --- 01 / 01b -------------------------------------------------------------
# Mapa de filas NUEVO de cada pestaña de año. Lo que cambia respecto de la
# v1.1: el food cost se parte en comida y bebida (RD-14) y aparece la línea de
# comisiones de delivery (RD-13), así que TOTAL GASTOS baja de la 19 a la 21 y
# EBITDA de la 21 a la 23. Los `Resumen` de más abajo leen ESTAS filas.
FILA_TOTAL_INGRESOS = 10
GASTOS_ANUAL = (
    (13, 'Coste de comida (materia prima)', 'input'),
    (14, 'Coste de bebida (barra)', 'input'),
    (15, 'Personal (Labor)', 'input'),
    (16, 'Comisiones de plataformas (delivery)', 'delivery'),
    (17, 'Alquiler', 'input'),
    (18, 'Suministros', 'input'),
    (19, 'Marketing', 'input'),
    (20, 'Administración / Otros', 'input'),
)
FILA_TOTAL_GASTOS = 21
FILA_EBITDA = 23
FILA_MARGEN = 24
FILA_PARAM_DELIVERY = 27

NOTA_PACKAGING = ('Envases y packaging del delivery: inclúyelos en «Coste de '
                  'comida» o en «Administración / Otros»; no son gratis.')

# --- 05 -------------------------------------------------------------------
GASTOS_MES = GASTOS_ANUAL[:3] + (
    (16, 'Comisiones de plataformas (delivery)', 'input'),
) + GASTOS_ANUAL[4:]
FILA_TOTAL_GASTOS_MES = 21
FILA_EBITDA_MES = 23
FILA_MARGEN_MES = 24
RATIOS_MES = (
    (27, 'Food Cost % (sobre ventas de comida)', 'menor'),
    (28, 'Beverage Cost % (sobre ventas de barra)', 'menor'),
    (29, 'Labor Cost %', 'menor'),
    (30, 'Prime Cost % (comida + bebida + personal)', 'menor'),
)
FILA_ULTIMA_MES = 30          # hasta aquí llega el semáforo (motor.semaforos)

# Tolerancias del semáforo, en celda (RT-25): 'Resumen Anual'!B22:B25.
TOL_PCT_VERDE = "'Resumen Anual'!$B$22"
TOL_PCT_AMBAR = "'Resumen Anual'!$B$23"
TOL_PP_VERDE = "'Resumen Anual'!$B$24"
TOL_PP_AMBAR = "'Resumen Anual'!$B$25"



def _es_num(x):
    """pycel devuelve `int` cuando el resultado es entero: comprobar sólo
    `float` daba falsos fallos (121000 no es `float`)."""
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def _sem_mayor_mejor(celda, verde_ref, ambar_ref):
    """Ingresos, EBITDA y márgenes: lo malo es quedarse CORTO.

    RX-01 — el `ISNUMBER` no es decorativo: la celda de desviación devuelve
    TEXTO cuando no hay presupuesto (cadena vacía de `_desviacion_pct`) o
    cuando un `IFERROR` se dispara, y en Excel `"" >= -0,05` es VERDADERO
    (cualquier texto gana a cualquier número). Sin la guarda, las 240 celdas
    de `F6:F30` de las 12 pestañas salían «✅ OK» con el libro vacío, y un
    EBITDA real de −8.000 € contra un presupuesto de 0 también.
    """
    return ('=IFERROR(IF(ISNUMBER(' + celda + '),IF(' + celda + '>=-'
            + verde_ref + ',"✅ OK",IF(' + celda + '>=-' + ambar_ref
            + ',"⚠️ Atención","🔴 Alerta")),"—"),"—")')


def _sem_menor_mejor(celda, verde_ref, ambar_ref):
    """Gastos y ratios de coste: lo malo es PASARSE. `ISNUMBER`: RX-01."""
    return ('=IFERROR(IF(ISNUMBER(' + celda + '),IF(' + celda + '<='
            + verde_ref + ',"✅ OK",IF(' + celda + '<=' + ambar_ref
            + ',"⚠️ Atención","🔴 Alerta")),"—"),"—")')


def _desviacion_pct(fila, divisor=None):
    """Desviación % de una fila del 05 (RX-05).

    «Sin presupuesto» y «desviación cero» NO son lo mismo. La fórmula anterior
    —`=IF(B6=0,0,(C6-B6)/B6)`— devolvía 0 con el presupuesto en blanco, que es
    el estado en que se entrega el fichero, y el semáforo lo leía como
    «clavado»: 240 verdes de fábrica. Ahora devuelve la cadena vacía y el
    `ISNUMBER` de los semáforos la traduce a «—».
    """
    f = str(fila)
    div = divisor or ('$B' + f)
    return ('=IF($B' + f + '=0,"",IFERROR(($C' + f + '-$B' + f + ')/' + div
            + ',""))')


def _col(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


def _clonar_estilo(ws, origen, destinos):
    est = copy.copy(ws[origen]._style)
    for d in destinos:
        ws[d]._style = copy.copy(est)


# ==========================================================================
# 01 y 01b — pestañas de año
# ==========================================================================
def _pestana_anual(ws, informe):
    """Reescribe el bloque de GASTOS hacia abajo con el mapa nuevo."""
    # Los estilos se copian SÓLO de celdas cuyo papel es el mismo antes y
    # después del mapa nuevo. Copiarlos de `A19`/`B19` —TOTAL GASTOS en la
    # v1.1, «Marketing» (input verde) en la v2.0— hacía que la 2.ª pasada
    # pintara de verde las fórmulas de total: 721 diferencias de idempotencia
    # que en Excel se habrían visto como celdas de resultado «editables».
    est_label = copy.copy(ws['A6']._style)
    est_input = copy.copy(ws['B6']._style)
    est_total_lbl = copy.copy(ws['A10']._style)
    est_total_val = copy.copy(ws['B10']._style)
    est_titulo = copy.copy(ws['A5']._style)

    motor.limpiar_rango(ws, 'A12:N40')

    ws['A12'].value = 'GASTOS'
    ws['A12']._style = copy.copy(est_titulo)

    for fila, etiqueta, tipo in GASTOS_ANUAL:
        cel = ws['A' + str(fila)]
        cel.value = etiqueta
        cel._style = copy.copy(est_label)
        for c in range(2, 14):                      # B..M
            coord = _col(c) + str(fila)
            if tipo == 'delivery':
                # 30-35 % de la venta de delivery, con el % en celda (§1.3).
                motor.aplicar_estilo(ws, coord, est_total_val)
                motor.f(ws, coord,
                        '=' + _col(c) + '8*$B$' + str(FILA_PARAM_DELIVERY),
                        motor.FMT_EUR)
            else:
                motor.aplicar_estilo(ws, coord, est_input)
                motor.val(ws, coord, 0, motor.FMT_EUR, verde_=True)
        motor.aplicar_estilo(ws, 'N' + str(fila), est_total_val)
        motor.f(ws, 'N' + str(fila),
                '=SUM(B' + str(fila) + ':M' + str(fila) + ')', motor.FMT_EUR)

    ws['A' + str(FILA_TOTAL_GASTOS)].value = 'TOTAL GASTOS'
    ws['A' + str(FILA_TOTAL_GASTOS)]._style = copy.copy(est_total_lbl)
    for c in range(2, 15):
        L = _col(c)
        motor.aplicar_estilo(ws, L + str(FILA_TOTAL_GASTOS), est_total_val)
        motor.f(ws, L + str(FILA_TOTAL_GASTOS),
                '=SUM(' + L + '13:' + L + '20)', motor.FMT_EUR)

    ws['A' + str(FILA_EBITDA)].value = 'EBITDA'
    ws['A' + str(FILA_EBITDA)]._style = copy.copy(est_total_lbl)
    ws['A' + str(FILA_MARGEN)].value = 'Margen EBITDA %'
    ws['A' + str(FILA_MARGEN)]._style = copy.copy(est_label)
    for c in range(2, 15):
        L = _col(c)
        motor.aplicar_estilo(ws, L + str(FILA_EBITDA), est_total_val)
        motor.f(ws, L + str(FILA_EBITDA),
                '=' + L + str(FILA_TOTAL_INGRESOS) + '-' + L
                + str(FILA_TOTAL_GASTOS), motor.FMT_EUR)
        motor.aplicar_estilo(ws, L + str(FILA_MARGEN), est_total_val,
                             motor.FMT_PCT)
        motor.f(ws, L + str(FILA_MARGEN),
                '=IF(' + L + str(FILA_TOTAL_INGRESOS) + '=0,0,' + L
                + str(FILA_EBITDA) + '/' + L + str(FILA_TOTAL_INGRESOS) + ')',
                motor.FMT_PCT)

    ws['A26'].value = 'PARÁMETROS Y NOTAS'
    ws['A26']._style = copy.copy(est_titulo)
    ws['A' + str(FILA_PARAM_DELIVERY)].value = (
        '% de comisión de las plataformas de delivery (sobre la venta de '
        'delivery de la fila 8)')
    ws['A' + str(FILA_PARAM_DELIVERY)]._style = copy.copy(est_label)
    motor.val(ws, 'B' + str(FILA_PARAM_DELIVERY), 0.30, motor.FMT_PCT,
              verde_=True)
    ws['A28'].value = NOTA_PACKAGING
    ws['A28'].font = Font(italic=True, size=9)
    ws['A30'].value = '© 2026 AI Chef Pro · aichef.pro'
    ws['A30'].font = Font(size=8)
    informe.append(ws.title + ': gastos con comisión de delivery y coste de '
                   'bebida (RD-13/RD-14); TOTAL en la fila '
                   + str(FILA_TOTAL_GASTOS) + ', EBITDA en la '
                   + str(FILA_EBITDA))


def _resumen_previsional(ws, anios, informe):
    """`01!Resumen` / `01b!Resumen`: consolidación POR REFERENCIA (DOM-04)."""
    ultima = 1 + anios                      # D con 3 años, F con 5
    crec = _col(ultima + 1)                 # E o G
    primera = 'B'
    ultima_l = _col(ultima)
    for i in range(anios):
        L = _col(2 + i)
        hoja = "'Año " + str(i + 1) + "'!"
        motor.f(ws, L + '5', '=' + hoja + 'N' + str(FILA_TOTAL_INGRESOS),
                motor.FMT_EUR)
        motor.f(ws, L + '6', '=' + hoja + 'N' + str(FILA_TOTAL_GASTOS),
                motor.FMT_EUR)
        motor.f(ws, L + '7', '=' + hoja + 'N' + str(FILA_EBITDA),
                motor.FMT_EUR)
        motor.f(ws, L + '8', motor.iferror('$' + L + '7/$' + L + '5'),
                motor.FMT_PCT)
    # Crecimiento Y1→Yn. El margen (fila 8) va en PUNTOS PORCENTUALES: un
    # margen no «crece un 20 %», sube o baja p.p. (DOM-27).
    for fila in (5, 6, 7):
        motor.f(ws, crec + str(fila),
                motor.iferror('($' + ultima_l + str(fila) + '-$' + primera
                              + str(fila) + ')/$' + primera + str(fila)),
                motor.FMT_PCT)
    # RX-02: B8 y D8/F8 son `=IFERROR($B7/$B5,"")` y con el libro en blanco
    # devuelven la cadena vacía; `"" - ""` es #¡VALOR!, y así se entregaba
    # cacheado en el Resumen de las dos plantillas estrella. Mismo helper que
    # las filas 5-7.
    motor.f(ws, crec + '8',
            motor.iferror('$' + ultima_l + '8-$' + primera + '8'),
            motor.FMT_PP)
    ws[crec + '4'].value = ('Crecimiento Y1→Y' + str(anios)
                            + ' (margen en p.p.)')
    informe.append('Resumen: ' + str(anios * 4)
                   + ' celdas consolidadas por referencia + columna de '
                     'crecimiento completa (DOM-04/DOM-27)')


# ==========================================================================
# 05 — pestañas mensuales y Resumen Anual
# ==========================================================================
def _pestana_mensual(ws, informe):
    est_label = copy.copy(ws['A13']._style)
    est_input = copy.copy(ws['B13']._style)
    est_calc = copy.copy(ws['D13']._style)
    est_pct = copy.copy(ws['E13']._style)
    est_sem = copy.copy(ws['F13']._style)
    # A19/B19 son TOTAL GASTOS en la v1.1 y «Marketing» en la v2.0: el estilo
    # se copia de la fila 10 (TOTAL INGRESOS), que no cambia de papel.
    est_total_lbl = copy.copy(ws['A10']._style)
    est_total_val = copy.copy(ws['B10']._style)
    est_titulo = copy.copy(ws['A12']._style)

    # 1) INGRESOS: sólo cambia el semáforo (deja de castigar vender de más).
    for fila in (6, 7, 8, 9, 10):
        motor.f(ws, 'E' + str(fila), _desviacion_pct(fila), motor.FMT_PCT)
        motor.f(ws, 'F' + str(fila),
                _sem_mayor_mejor('$E' + str(fila), TOL_PCT_VERDE,
                                 TOL_PCT_AMBAR))

    # 2) GASTOS hacia abajo: mapa nuevo.
    motor.limpiar_rango(ws, 'A12:F40')
    ws['A12'].value = 'GASTOS'
    ws['A12']._style = copy.copy(est_titulo)

    for fila, etiqueta, _tipo in GASTOS_MES:
        r = str(fila)
        ws['A' + r].value = etiqueta
        ws['A' + r]._style = copy.copy(est_label)
        for coord in ('B' + r, 'C' + r):
            motor.aplicar_estilo(ws, coord, est_input)
            motor.val(ws, coord, 0, motor.FMT_EUR, verde_=True)
        motor.aplicar_estilo(ws, 'D' + r, est_calc, motor.FMT_EUR)
        motor.f(ws, 'D' + r, '=C' + r + '-B' + r, motor.FMT_EUR)
        motor.aplicar_estilo(ws, 'E' + r, est_pct, motor.FMT_PCT)
        motor.f(ws, 'E' + r, _desviacion_pct(fila), motor.FMT_PCT)
        motor.aplicar_estilo(ws, 'F' + r, est_sem)
        motor.f(ws, 'F' + r,
                _sem_menor_mejor('$E' + r, TOL_PCT_VERDE, TOL_PCT_AMBAR))

    r = str(FILA_TOTAL_GASTOS_MES)
    ws['A' + r].value = 'TOTAL GASTOS'
    ws['A' + r]._style = copy.copy(est_total_lbl)
    for L in ('B', 'C', 'D', 'E', 'F'):
        motor.aplicar_estilo(ws, L + r, est_total_val)
    for L in ('B', 'C'):
        motor.f(ws, L + r, '=SUM(' + L + '13:' + L + '20)', motor.FMT_EUR)
    motor.f(ws, 'D' + r, '=C' + r + '-B' + r, motor.FMT_EUR)
    motor.f(ws, 'E' + r, _desviacion_pct(FILA_TOTAL_GASTOS_MES),
            motor.FMT_PCT)
    motor.f(ws, 'F' + r,
            _sem_menor_mejor('$E' + r, TOL_PCT_VERDE, TOL_PCT_AMBAR))

    # EBITDA: la desviación % divide entre ABS() del presupuesto (TEC-10).
    e = str(FILA_EBITDA_MES)
    ws['A' + e].value = 'EBITDA'
    ws['A' + e]._style = copy.copy(est_total_lbl)
    for L in ('B', 'C', 'D', 'E', 'F'):
        motor.aplicar_estilo(ws, L + e, est_total_val)
    for L in ('B', 'C'):
        motor.f(ws, L + e, '=' + L + '10-' + L + r, motor.FMT_EUR)
    motor.f(ws, 'D' + e, '=C' + e + '-B' + e, motor.FMT_EUR)
    motor.f(ws, 'E' + e,
            _desviacion_pct(FILA_EBITDA_MES, 'ABS($B' + e + ')'),
            motor.FMT_PCT)
    motor.f(ws, 'F' + e,
            _sem_mayor_mejor('$E' + e, TOL_PCT_VERDE, TOL_PCT_AMBAR))

    # Margen EBITDA % — fila que hasta hoy sólo tenía B y C (TEC-26).
    m = str(FILA_MARGEN_MES)
    ws['A' + m].value = 'Margen EBITDA %'
    ws['A' + m]._style = copy.copy(est_label)
    for L in ('B', 'C'):
        motor.aplicar_estilo(ws, L + m, est_pct, motor.FMT_PCT)
        # RX-05: sin ventas el margen no es 0 %, es desconocido. Devolver 0
        # hacía que la diferencia en p.p. fuese 0 y el semáforo verde.
        motor.f(ws, L + m, '=IF(' + L + '10=0,"",' + L + e + '/' + L + '10)',
                motor.FMT_PCT)
    motor.aplicar_estilo(ws, 'D' + m, est_pct, motor.FMT_PP)
    motor.f(ws, 'D' + m, motor.iferror('$C' + m + '-$B' + m), motor.FMT_PP)
    ws['E' + m].value = 'p.p.'
    ws['E' + m].alignment = Alignment(horizontal='center')
    motor.aplicar_estilo(ws, 'F' + m, est_sem)
    motor.f(ws, 'F' + m,
            _sem_mayor_mejor('$D' + m, TOL_PP_VERDE, TOL_PP_AMBAR))

    # RATIOS AUTOMÁTICOS, ahora con D/E/F (TEC-26) y con el food cost medido
    # sobre las ventas de COMIDA, no sobre el total con la barra dentro.
    ws['A26'].value = 'RATIOS AUTOMÁTICOS'
    ws['A26']._style = copy.copy(est_titulo)
    # RX-05: la rama «sin base» devuelve «» (no 0), para que la diferencia
    # en p.p. de la columna D no valga 0 y el semáforo no se ponga verde.
    formulas = {
        27: '=IF(({L}10-{L}7)<=0,"",{L}13/({L}10-{L}7))',
        28: '=IF({L}7=0,"",{L}14/{L}7)',
        29: '=IF({L}10=0,"",{L}15/{L}10)',
        30: '=IF({L}10=0,"",({L}13+{L}14+{L}15)/{L}10)',
    }
    for fila, etiqueta, sentido in RATIOS_MES:
        rr = str(fila)
        ws['A' + rr].value = etiqueta
        ws['A' + rr]._style = copy.copy(est_label)
        for L in ('B', 'C'):
            motor.aplicar_estilo(ws, L + rr, est_pct, motor.FMT_PCT)
            motor.f(ws, L + rr, formulas[fila].replace('{L}', L),
                    motor.FMT_PCT)
        motor.aplicar_estilo(ws, 'D' + rr, est_pct, motor.FMT_PP)
        motor.f(ws, 'D' + rr, motor.iferror('$C' + rr + '-$B' + rr),
                motor.FMT_PP)
        ws['E' + rr].value = 'p.p.'
        ws['E' + rr].alignment = Alignment(horizontal='center')
        motor.aplicar_estilo(ws, 'F' + rr, est_sem)
        motor.f(ws, 'F' + rr,
                _sem_menor_mejor('$D' + rr, TOL_PP_VERDE, TOL_PP_AMBAR)
                if sentido == 'menor'
                else _sem_mayor_mejor('$D' + rr, TOL_PP_VERDE, TOL_PP_AMBAR))

    ws['A32'].value = '© 2026 AI Chef Pro · aichef.pro'
    ws['A32'].font = Font(size=8)
    informe.append(ws.title + ': semáforo con signo, desviaciones en p.p. y '
                              'ratios completos (DOM-12/TEC-09/TEC-10/TEC-26)')


def _resumen_anual(ws, informe):
    est_label = copy.copy(ws['A5']._style)
    est_val = copy.copy(ws['B5']._style)

    motor.limpiar_rango(ws, 'A11:N45')

    for i, mes in enumerate(MESES):
        L = _col(2 + i)
        h = "'" + mes + "'!"
        motor.f(ws, L + '5', '=' + h + 'B10', motor.FMT_EUR)
        motor.f(ws, L + '6', '=' + h + 'C10', motor.FMT_EUR)
        motor.f(ws, L + '7', '=' + h + 'B' + str(FILA_TOTAL_GASTOS_MES),
                motor.FMT_EUR)
        motor.f(ws, L + '8', '=' + h + 'C' + str(FILA_TOTAL_GASTOS_MES),
                motor.FMT_EUR)
        motor.f(ws, L + '9', '=' + h + 'B' + str(FILA_EBITDA_MES),
                motor.FMT_EUR)
        motor.f(ws, L + '10', '=' + h + 'C' + str(FILA_EBITDA_MES),
                motor.FMT_EUR)
        # Filas nuevas de §2: desviaciones anuales y coste de comida/bebida.
        motor.f(ws, L + '11', '=' + L + '6-' + L + '5', motor.FMT_EUR)
        motor.f(ws, L + '12', motor.iferror('(' + L + '6-' + L + '5)/ABS('
                                            + L + '5)'), motor.FMT_PCT)
        motor.f(ws, L + '13', '=' + L + '10-' + L + '9', motor.FMT_EUR)
        motor.f(ws, L + '14', '=' + h + 'C13', motor.FMT_EUR)
        motor.f(ws, L + '15', '=' + h + 'C14', motor.FMT_EUR)
        motor.f(ws, L + '16', '=' + h + 'C15', motor.FMT_EUR)
        motor.f(ws, L + '17', '=' + L + '14+' + L + '15+' + L + '16',
                motor.FMT_EUR)
        motor.f(ws, L + '18', '=IF(' + L + '6=0,0,' + L + '17/' + L + '6)',
                motor.FMT_PCT)

    etiquetas = {
        11: 'Desviación de ingresos (€)',
        12: 'Desviación de ingresos (%)',
        13: 'Desviación de EBITDA (€)',
        14: 'Coste de comida (real, €)',
        15: 'Coste de bebida (real, €)',
        16: 'Personal — Labor (real, €)',
        17: 'Prime Cost (real, €)',
        18: 'Prime Cost % (real)',
    }
    for fila, texto in etiquetas.items():
        ws['A' + str(fila)].value = texto
        ws['A' + str(fila)]._style = copy.copy(est_label)

    # Columna TOTAL — hoy tampoco sumaba (DOM-05/RD-08).
    for fila in range(5, 18):
        if fila == 12:
            motor.f(ws, 'N12', motor.iferror('(N6-N5)/ABS(N5)'),
                    motor.FMT_PCT)
            continue
        motor.aplicar_estilo(ws, 'N' + str(fila), est_val, motor.FMT_EUR)
        motor.f(ws, 'N' + str(fila), '=SUM(B' + str(fila) + ':M' + str(fila)
                + ')', motor.FMT_EUR)
    motor.f(ws, 'N18', '=IF(N6=0,0,N17/N6)', motor.FMT_PCT)

    # Tolerancias del semáforo EN CELDA (RT-25): las leen las 12 pestañas.
    ws['A21'].value = 'TOLERANCIAS DEL SEMÁFORO (edítalas si tu negocio manda)'
    ws['A21'].font = Font(bold=True)
    tolerancias = (
        (22, 'Desviación tolerada — verde (± % sobre presupuesto)', 0.05,
         motor.FMT_PCT),
        (23, 'Desviación de atención — ámbar (± %)', 0.10, motor.FMT_PCT),
        (24, 'Desviación tolerada en ratios — verde (p.p.)', 0.02,
         motor.FMT_PP),
        (25, 'Desviación de atención en ratios — ámbar (p.p.)', 0.05,
         motor.FMT_PP),
    )
    for fila, texto, valor, fmt in tolerancias:
        ws['A' + str(fila)].value = texto
        ws['A' + str(fila)]._style = copy.copy(est_label)
        motor.val(ws, 'B' + str(fila), valor, fmt, verde_=True)

    ws['A45'].value = '© 2026 AI Chef Pro · aichef.pro'
    ws['A45'].font = Font(size=8)
    motor.anchos(ws, {'A': 46})
    informe.append("'Resumen Anual': 12 meses consolidados por referencia, "
                   'columna TOTAL que suma y tolerancias del semáforo en '
                   'celda (DOM-05/TEC-04/RT-25)')


# ==========================================================================
# API del grupo
# ==========================================================================
def post(wb, fname, cambios, registro):
    if fname.startswith('01'):
        anios = 5 if fname.startswith('01b') else 3
        for i in range(anios):
            hoja = 'Año ' + str(i + 1)
            if hoja in wb.sheetnames:
                _pestana_anual(wb[hoja], cambios)
        if 'Resumen' in wb.sheetnames:
            _resumen_previsional(wb['Resumen'], anios, cambios)
        if 'Instrucciones' in wb.sheetnames:
            ws = wb['Instrucciones']
            motor.linea_instrucciones(
                ws, '▸ El gasto de delivery lleva su propia línea de comisión '
                    'de plataformas: el % está en la fila '
                    + str(FILA_PARAM_DELIVERY) + ' de cada pestaña de año.')
            motor.linea_instrucciones(
                ws, '▸ El coste de materia prima va desdoblado en comida y '
                    'bebida, para que el Food Cost % no se diluya con la '
                    'barra dentro.')
    elif fname.startswith('05'):
        for mes in MESES:
            if mes in wb.sheetnames:
                _pestana_mensual(wb[mes], cambios)
        if 'Resumen Anual' in wb.sheetnames:
            _resumen_anual(wb['Resumen Anual'], cambios)
        if 'Instrucciones' in wb.sheetnames:
            ws = wb['Instrucciones']
            motor.linea_instrucciones(
                ws, '▸ El semáforo va CON SIGNO: en ingresos y EBITDA avisa '
                    'cuando te quedas corto; en gastos, cuando te pasas. Los '
                    'umbrales son editables en «Resumen Anual» (B22:B25).',
                rx=__import__('re').compile(r'^▸ El semáforo automático'))
            motor.linea_instrucciones(
                ws, '▸ Food Cost % se mide sobre las ventas de comida (total '
                    'menos barra) y Beverage Cost % sobre las de barra.',
                rx=__import__('re').compile(r'^▸ Food Cost %'))
    return cambios


def demos(carpeta, origen, destino):
    """Pruebas con pycel de lo que este grupo promete (§5 de la SPEC)."""
    import contextlib
    import os
    import shutil

    fuera = {'fallos': [], 'grupo_a': {}}

    def _ev(xl, ref):
        with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
            try:
                return xl.evaluate(ref)
            except Exception as e:                           # noqa: BLE001
                return 'ERR:' + type(e).__name__

    os.makedirs(destino, exist_ok=True)
    from pycel import ExcelCompiler

    # --- 05: el semáforo deja de castigar vender por encima del presupuesto.
    f05 = '05-pyl-mensual-real-vs-presupuesto.xlsx'
    p05 = os.path.join(carpeta, f05)
    if os.path.isfile(p05):
        copia = os.path.join(destino, '_ga_' + f05)
        shutil.copy2(p05, copia)
        xl = ExcelCompiler(filename=copia)
        # OJO al ORDEN: hay que evaluar la celda de DESVIACIÓN antes que la
        # del semáforo. Si se compila primero `F6`, pycel no invalida `E6` al
        # hacer `set_value` sobre `B6`/`C6` y el semáforo se queda con el «—»
        # de la primera pasada — un falso fallo que no existe en Excel
        # (comprobado: evaluando `E6` primero, `F6` da «✅ OK»).
        sin_ppto_ingresos = (_ev(xl, "'Ene'!E6"), _ev(xl, "'Ene'!F6"))
        xl.set_value("'Ene'!B6", 10000)
        xl.set_value("'Ene'!C6", 12000)
        vender_mas = _ev(xl, "'Ene'!F6")
        sin_ppto_gastos = (_ev(xl, "'Ene'!E13"), _ev(xl, "'Ene'!F13"))
        xl.set_value("'Ene'!B13", 10000)
        xl.set_value("'Ene'!C13", 12000)
        gastar_mas = _ev(xl, "'Ene'!F13")
        # EBITDA presupuestado NEGATIVO: una mejora tiene que salir en verde.
        xl2 = ExcelCompiler(filename=copia)
        _ev(xl2, "'Ene'!E" + str(FILA_EBITDA_MES))
        _ev(xl2, "'Ene'!F" + str(FILA_EBITDA_MES))
        xl2.set_value("'Ene'!B" + str(FILA_EBITDA_MES), -3000)
        xl2.set_value("'Ene'!C" + str(FILA_EBITDA_MES), 0)
        ebitda_neg = _ev(xl2, "'Ene'!F" + str(FILA_EBITDA_MES))
        desv_neg = _ev(xl2, "'Ene'!E" + str(FILA_EBITDA_MES))
        # Consolidación del Resumen Anual.
        xl3 = ExcelCompiler(filename=copia)
        _ev(xl3, "'Resumen Anual'!B6")
        _ev(xl3, "'Ene'!C10")
        xl3.set_value("'Ene'!C6", 25000)
        consolida = _ev(xl3, "'Resumen Anual'!B6")

        d = {
            'sin_presupuesto_ingresos_E6_F6': sin_ppto_ingresos,
            'sin_presupuesto_gastos_E13_F13': sin_ppto_gastos,
            'vender_20pct_por_encima': vender_mas,
            'gastar_20pct_de_mas': gastar_mas,
            'EBITDA_ppto_-3000_real_0': ebitda_neg,
            'desviacion_con_ABS_en_denominador': desv_neg,
            "'Resumen Anual'!B6_con_'Ene'!C6=25000": consolida,
        }
        fuera['grupo_a']['05'] = d
        # RX-01/RX-05: sin presupuesto el semáforo NO opina. Antes salían
        # 240 «✅ OK» de fábrica en las 12 pestañas.
        for etiqueta, par in (('ingresos', sin_ppto_ingresos),
                              ('gastos', sin_ppto_gastos)):
            if par[1] != '—':
                fuera['fallos'].append(
                    '05: sin presupuesto, el semáforo de ' + etiqueta
                    + ' no calla (' + str(par) + ')')
        if not (isinstance(vender_mas, str) and 'OK' in vender_mas):
            fuera['fallos'].append(
                '05: vender un 20 % por encima del presupuesto sigue sin ser '
                'OK (' + str(vender_mas) + ')')
        if not (isinstance(gastar_mas, str) and 'Alerta' in gastar_mas):
            fuera['fallos'].append(
                '05: gastar un 20 % de más ya no salta (' + str(gastar_mas)
                + ')')
        if not (isinstance(ebitda_neg, str) and 'OK' in ebitda_neg):
            fuera['fallos'].append(
                '05: con EBITDA presupuestado negativo, una mejora sigue sin '
                'salir en verde (' + str(ebitda_neg) + ')')
        if consolida != 25000:
            fuera['fallos'].append(
                "05: 'Resumen Anual'!B6 no consolida 'Ene'!C10 ("
                + str(consolida) + ')')

    # --- 01: el Resumen consolida por referencia.
    f01 = '01-plan-financiero-previsional.xlsx'
    p01 = os.path.join(carpeta, f01)
    if os.path.isfile(p01):
        copia = os.path.join(destino, '_ga_' + f01)
        shutil.copy2(p01, copia)
        xl = ExcelCompiler(filename=copia)
        _ev(xl, "'Resumen'!B5")
        _ev(xl, "'Año 1'!N10")
        xl.set_value("'Año 1'!B6", 30000)
        b5 = _ev(xl, "'Resumen'!B5")
        # Comisión de delivery: 30 % de la venta de delivery del mes.
        xl2 = ExcelCompiler(filename=copia)
        _ev(xl2, "'Año 1'!B16")
        xl2.set_value("'Año 1'!B8", 10000)
        comision = _ev(xl2, "'Año 1'!B16")
        fuera['grupo_a']['01'] = {
            "'Resumen'!B5_con_'Año 1'!B6=30000": b5,
            'comision_delivery_sobre_10000': comision,
        }
        if b5 != 30000:
            fuera['fallos'].append(
                "01: 'Resumen'!B5 no consolida 'Año 1'!N10 (" + str(b5) + ')')
        if not (_es_num(comision) and abs(comision - 3000) < 0.01):
            fuera['fallos'].append(
                '01: la comisión de delivery no calcula el 30 % ('
                + str(comision) + ')')
    return fuera
