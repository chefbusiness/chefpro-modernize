#!/usr/bin/env python3
"""
grupo_c.py — §4 de la SPEC: informe bancario, ratios y bonus (06, 07,
BONUS-08, BONUS-09).

Qué arregla, id por id:

  · DOM-02 / TEC-02 / COM-18 / RD-03 / RT-06 (alta) — el 07, que se vende
    «para presentarlo a un banco», salía con **40 ceros constantes** y tres
    fórmulas en todo el libro: TOTAL GASTOS, EBITDA, BAI, Impuesto, BENEFICIO
    NETO y CASH FLOW OPERATIVO eran constantes SIN relleno de input, así que se
    leen como resultados. Ahora el P&L calcula línea a línea.
  · DOM-03 / TEC-03 / COM-02 / RD-04 / RC-02 / RC-20 (alta) — `IRR`/`NPV`
    operaban sobre `B19:F19`, donde sólo existía `B19=-150000`: la TIR salía
    «—» y el VAN se entregaba como −150.000,00 €. Y el rango cogía **cuatro**
    años de cinco. Bloque de flujos nuevo con columna de AÑO 0 (`B22:G22`),
    flujo acumulado, TIR (cacheada por Newton, §1.8), VAN con la tasa **en
    celda** y payback por fórmula evaluable.
  · DOM-11 / TEC-19 / COM-17 / RD-05 / RC-03 — `07!Ratios!C4:C10` no tenía ni
    celdas creadas, y ahí está el DSCR, que es el número que decide la
    operación. Se calculan los siete contra un bloque de balance y deuda nuevo.
  · DOM-14 / RD-05 — hoja `Financiación` con cuadro de amortización por
    **anualidad algebraica** (`PMT` está prohibido: pycel no lo implementa y su
    `IFERROR` no lo atrapa). De ahí salen los intereses del P&L, la cuota del
    DSCR y la fila 23 del 03.
  · DOM-19 / RC-27 — el 25 % del Impuesto sobre Sociedades iba clavado en la
    ETIQUETA, en un documento cuyo destinatario típico es una entidad de nueva
    creación (15 % los dos primeros ejercicios con base positiva, art. 29.1
    LIS) y que puede ni siquiera ser una SL. Pasa a celda con su nota (§7.6).
  · DOM-21 / COM-16 / RD-21 / RT-07 / RT-21 — los 150.000 € de inversión
    ficticia dejan de estar precargados: `B22` sale del Resumen Ejecutivo. Y
    como ya no es un input numérico, deja de chocar con la validación «≥ 0»
    que el propio motor le ponía encima (el fichero se entregaba violando su
    propia validación).
  · DOM-07 / TEC-13 / COM-07 / RD-11 / RC-15 — el RevPASH dividía por METROS
    CUADRADOS y se comparaba con un umbral por PLAZA: el ejemplo daba 2,95 €
    frente a «> 8 €» y el comprador concluía que iba tres veces por debajo del
    sector. Input de aforo nuevo, y las ventas por m² pasan a fila propia.
  · DOM-18 / TEC-14 / COM-32 — «Gastos operativos totales» no definía qué
    incluye y en USALI el GOP va antes de rentas y cargas de propiedad: dos
    usuarios con el mismo negocio obtenían GOP separados por 10-15 puntos.
    Pasa a CALCULADO y el alquiler sale a su propia línea, con EBITDA % aparte.
  · DOM-29 / TEC-12 / COM-29 / DOM-22 / RD-19 / RD-20 / RT-09 / RC-16 — el
    food cost valía 32 % en las Instrucciones y en la columna de referencia y
    33 % en la fórmula: un 30 %, que el propio manual llama ideal, salía
    «ATENCIÓN Aceptable». Los umbrales viven ahora en columnas NUMÉRICAS de
    `Benchmarks` y **todo lo demás se genera desde ahí**: el texto de la
    columna «Óptimo/Aceptable/Peligro», la referencia que se enseña en Ratios,
    la fórmula del semáforo y las Instrucciones. No pueden divergir porque hay
    una sola fuente.
  · DOM-06 / TEC-21 / TEC-22 / COM-08 / COM-26 / RD-10 / RT-20 / RC-11 —
    BONUS-08 contaba el personal dos veces (19.000 € de «costes fijos» que ya
    incluían 12.000 de nóminas, más un 28 % de labor cost encima): el escenario
    BASE, el que se enseña a un inversor, cerraba en −5.272 €/mes y −63.264
    €/año. Costes fijos SIN personal, base común con el 02 y escenarios
    recalibrados.
  · DOM-20 / DOM-26 / TEC-24 / COM-30 / RD-07 / RT-08 / RC-17 — la checklist
    financiera pre-apertura no tenía **ni una** tarea de alta laboral, todas
    con sanción directa e impacto de caja el primer mes, mientras la fase 6
    detallaba la «política de propinas». Fase 7 con 6 tareas → 54. Contador
    honesto (`COUNTIF`, no el 48 escrito a mano), DV con
    `showErrorMessage=True` sobre `F5:F64` y —lo que nadie había visto— la
    columna Estado DESBLOQUEADA: era la única que el usuario tiene que tocar y
    la protección la cerraba, mientras la banda de color de la fase 3 se abría
    por accidente por usar el mismo verde que los editables.
"""
import copy

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

import motor

FICHEROS = [
    '06-dashboard-ratios-financieros.xlsx',
    '07-informe-viabilidad-bancos.xlsx',
    'BONUS-08-simulador-escenarios.xlsx',
    'BONUS-09-checklist-pre-apertura.xlsx',
]

# ==========================================================================
# 06 · Dashboard de ratios
# ==========================================================================
# (fila, etiqueta, óptimo, límite, sentido). «menor» = por debajo del óptimo
# es excelente; «mayor» = por encima del óptimo es excelente. Es la ÚNICA
# fuente de umbrales del kit: de aquí salen los textos de las tres columnas de
# `Benchmarks`, la referencia que se enseña en `Ratios`, la fórmula del
# semáforo y las Instrucciones (DOM-29).
BENCHMARKS = (
    (4, 'Food Cost %', 0.28, 0.32, 'menor', 'pct'),
    (5, 'Labor Cost %', 0.25, 0.30, 'menor', 'pct'),
    (6, 'Prime Cost %', 0.60, 0.65, 'menor', 'pct'),
    (7, 'GOP %', 0.20, 0.15, 'mayor', 'pct'),
    (8, 'Alquiler / Ventas', 0.08, 0.12, 'menor', 'pct'),
    (9, 'Marketing / Ventas', 0.05, 0.08, 'menor', 'pct'),
    (10, 'EBITDA %', 0.15, 0.10, 'mayor', 'pct'),
    # Se deriva del GOP, no se inventa: si el GOP óptimo es > 20 % y el
    # peligro < 15 %, el coste operativo por cubierto sobre el ticket es su
    # complementario (< 80 % óptimo, > 85 % peligro). Con el « < 60 % » de
    # la v1.1 el propio ejemplo del kit nacía en rojo contra un umbral que
    # contradecía la fila del GOP de la misma tabla.
    (11, 'Coste cubierto / Ticket', 0.80, 0.85, 'menor', 'pct'),
    (12, 'Beverage Cost %', 0.18, 0.24, 'menor', 'pct'),
    (13, 'RevPASH (€/plaza/hora)', 6.0, 3.0, 'mayor', 'eur'),
    (14, 'Margen bruto %', 0.70, 0.65, 'mayor', 'pct'),
)
BM_POR_ETIQUETA = dict((b[1], b) for b in BENCHMARKS)

# (fila de Ratios, etiqueta, fórmula, formato, etiqueta de Benchmarks)
RATIOS_06 = (
    # RX-03: el IFERROR sólo atrapa el #¡DIV/0!. Si el cliente borra C6
    # (ventas totales) y deja C7 (barra) del ejemplo, el denominador es
    # NEGATIVO y no hay error: C17 daba −0,9 y el semáforo «✅ Excelente»
    # con un food cost del −90 %. Se anula EN ORIGEN, que además protege a
    # cualquier consumidor de C17, no sólo al semáforo.
    (17, 'Food Cost % (sobre ventas de comida)',
     '=IF(($C$6-$C$7)<=0,"Indica las ventas",$C$8/($C$6-$C$7))', 'pct',
     'Food Cost %'),
    (18, 'Labor Cost %', '=IFERROR($C$10/$C$6,"Indica las ventas")', 'pct',
     'Labor Cost %'),
    (19, 'Prime Cost % (comida + bebida + personal)',
     '=IFERROR(($C$8+$C$36+$C$10)/$C$6,"Indica las ventas")', 'pct',
     'Prime Cost %'),
    (20, 'GOP (Gross Operating Profit) %',
     '=IFERROR(($C$6-$C$9)/$C$6,"Indica las ventas")', 'pct', 'GOP %'),
    (21, 'Beverage Cost % (sobre ventas de barra)',
     '=IFERROR($C$36/$C$7,"Indica las ventas de barra")', 'pct',
     'Beverage Cost %'),
    (22, 'RevPASH (€/plaza/hora)',
     '=IFERROR($C$6/($C$13*$C$12),"Indica plazas y horas")', 'eur',
     'RevPASH (€/plaza/hora)'),
    (23, 'EBITDA %',
     '=IFERROR(($C$6-$C$9-$C$38)/$C$6,"Indica las ventas")', 'pct',
     'EBITDA %'),
    (24, 'Coste por cubierto / Ticket medio',
     '=IFERROR($C$28/$C$29,"Indica los cubiertos")', 'pct',
     'Coste cubierto / Ticket'),
    (25, 'Margen bruto %',
     '=IFERROR(($C$6-$C$8-$C$36)/$C$6,"Indica las ventas")', 'pct',
     'Margen bruto %'),
)

# §1.4/§7.5 — el ejemplo del 06 tenía que contar la MISMA historia que el 02 y
# el BONUS-08. Ventas 31.460 € (22 € × 55 cubiertos × 26 días).
EJ = motor.EJEMPLO_CANONICO
VENTAS_EJ = EJ['ventas_mes']
BARRA_EJ = round(VENTAS_EJ * 0.25, 2)
COMIDA_EJ = round((VENTAS_EJ - BARRA_EJ) * EJ['food_cost_pct'], 2)
BEBIDA_EJ = round(BARRA_EJ * 0.22, 2)
PERSONAL_EJ = round(VENTAS_EJ * EJ['labor_cost_pct'], 2)
CUBIERTOS_EJ = EJ['cubiertos_dia'] * EJ['dias_apertura']



def _es_num(x):
    """pycel devuelve `int` cuando el resultado es entero: comprobar sólo
    `float` daba falsos fallos (121000 no es `float`)."""
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def _pct(v):
    return str(int(round(v * 100))) + '%'


def _num(v, tipo):
    return _pct(v) if tipo == 'pct' else str(int(round(v))) + '€'


def _texto_benchmark(opt, lim, sentido, tipo):
    """Los tres textos de `Benchmarks`, generados desde los NÚMEROS."""
    if sentido == 'menor':
        return ('< ' + _num(opt, tipo),
                _num(opt, tipo) + ' - ' + _num(lim, tipo),
                '> ' + _num(lim, tipo))
    return ('> ' + _num(opt, tipo),
            _num(lim, tipo) + ' - ' + _num(opt, tipo),
            '< ' + _num(lim, tipo))


def _benchmarks_06(wb, informe):
    ws = wb['Benchmarks']
    motor.limpiar_rango(ws, 'A3:H20')
    cabecera = ('Ratio', 'Óptimo', 'Aceptable', 'Peligro', 'Óptimo (nº)',
                'Límite (nº)', 'Sentido')
    for i, texto in enumerate(cabecera):
        cel = ws.cell(row=3, column=2 + i)
        cel.value = texto
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center', wrap_text=True)
    for fila, etiqueta, opt, lim, sentido, tipo in BENCHMARKS:
        r = str(fila)
        ws['B' + r] = etiqueta
        c, d, e = _texto_benchmark(opt, lim, sentido, tipo)
        for coord, texto, color in (('C' + r, c, 'C8E6C9'),
                                    ('D' + r, d, 'FFF9C4'),
                                    ('E' + r, e, 'FFCDD2')):
            ws[coord] = texto
            ws[coord].fill = motor.PatternFill('solid', fgColor=color)
            ws[coord].alignment = Alignment(horizontal='center')
        fmt = motor.FMT_PCT if tipo == 'pct' else motor.FMT_EUR
        motor.val(ws, 'F' + r, opt, fmt)
        motor.val(ws, 'G' + r, lim, fmt)
        ws['H' + r] = ('menor es mejor' if sentido == 'menor'
                       else 'mayor es mejor')
        ws['H' + r].font = Font(size=9, italic=True)
    ws['B17'] = ('Las columnas «Óptimo (nº)» y «Límite (nº)» son la ÚNICA '
                 'fuente de umbrales del kit: el semáforo de la pestaña '
                 'Ratios las lee, y los textos de la izquierda se generan '
                 'desde ellas. Edítalas y todo el dashboard se mueve con '
                 'ellas.')
    ws['B17'].font = Font(size=9, italic=True)
    ws['B17'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A20'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A20'].font = Font(size=8)
    motor.anchos(ws, {'B': 30, 'C': 14, 'D': 16, 'E': 14, 'F': 14, 'G': 14,
                      'H': 18})
    informe.append('06!Benchmarks: umbrales NUMÉRICOS en F/G y textos '
                   'generados desde ellos — una sola fuente (DOM-29/TEC-12)')


DATOS_06 = (
    (6, 'Ventas totales (€, sin IVA)', VENTAS_EJ, 'eur', 'Mensual'),
    (7, '· de las cuales, barra / bebida (€, sin IVA)', BARRA_EJ, 'eur',
     'Mensual'),
    (8, 'Coste de comida — consumo (€)', None, 'calc', 'Mensual'),
    (9, 'Gastos operativos totales (€)', None, 'calc', 'Mensual'),
    (10, 'Coste personal total (€)', PERSONAL_EJ, 'eur', 'Mensual'),
    (11, 'm² de sala', EJ['m2_sala'], 'ent', ''),
    (12, 'Horas de SERVICIO / mes (con clientes sentados)',
     EJ['horas_servicio_mes'], 'ent', ''),
    (13, 'Nº de plazas (aforo)', EJ['aforo_plazas'], 'ent', ''),
    (14, 'Nº de cubiertos servidos', CUBIERTOS_EJ, 'ent', 'Mensual'),
)

COMPLEMENTARIOS_06 = (
    (33, 'Existencias iniciales de comida (€)', 0.0,
     'Lo que había en despensa y cámara el día 1.'),
    (34, 'Compras de comida del período (€)', COMIDA_EJ,
     'Facturas de alimentación del período.'),
    (35, 'Existencias finales de comida (€)', 0.0,
     'Consumo = iniciales + compras − finales. El mes del pedido grande, las '
     'compras se disparan y el consumo NO.'),
    (36, 'Coste de bebida — consumo (€)', BEBIDA_EJ,
     'Barra y bodega, aparte de la comida.'),
    (37, 'Otros gastos operativos (€)', 5400.0,
     'Suministros, marketing, gestoría, comisiones de delivery… SIN alquiler '
     'ni amortización (van aparte, USALI). Con este valor el EBITDA del '
     'ejemplo sale al 17,3 %, el mismo que dan el 02 y el BONUS-08.'),
    (38, 'Alquiler (€)', 3000.0,
     'Fuera del GOP: en USALI es carga de propiedad.'),
)

OTROS_06 = (
    (28, 'Coste por cubierto (€)',
     '=IFERROR($C$9/$C$14,"Indica los cubiertos")', 'eur'),
    (29, 'Ticket medio (€, sin IVA)',
     '=IFERROR($C$6/$C$14,"Indica los cubiertos")', 'eur'),
    (30, 'Ventas por m² de sala (€)',
     '=IFERROR($C$6/$C$11,"Indica los m² de sala")', 'eur'),
)


def _ratios_06(wb, informe):
    ws = wb['Ratios']
    est_label = copy.copy(ws['B6']._style)
    est_input = copy.copy(ws['C6']._style)
    est_titulo = copy.copy(ws['B4']._style)

    motor.limpiar_rango(ws, 'A6:H50')

    fmts = {'eur': motor.FMT_EUR, 'ent': motor.FMT_ENT,
            'pct': motor.FMT_PCT}
    for fila, etiqueta, valor, tipo, periodo in DATOS_06:
        r = str(fila)
        ws['B' + r] = etiqueta
        ws['B' + r]._style = copy.copy(est_label)
        if tipo == 'calc':
            continue
        motor.aplicar_estilo(ws, 'C' + r, est_input, fmts[tipo])
        motor.val(ws, 'C' + r, valor, fmts[tipo], verde_=True)
        if periodo:
            ws['D' + r] = periodo
    # Consumo, no compras (RD-15), y GOP con alcance definido (DOM-18).
    motor.f(ws, 'C8', '=IFERROR($C$33+$C$34-$C$35,0)', motor.FMT_EUR)
    motor.f(ws, 'C9', '=IFERROR($C$8+$C$36+$C$10+$C$37,0)', motor.FMT_EUR)
    ws['D8'] = 'Calculado'
    ws['D9'] = 'Calculado'

    ws['B15'] = 'RATIOS CALCULADOS'
    ws['B15']._style = copy.copy(est_titulo)
    for i, texto in enumerate(('Ratio', 'Tu Valor', 'Benchmark Sector',
                               'Estado', 'Objetivo')):
        cel = ws.cell(row=16, column=2 + i)
        cel.value = texto
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center')

    for fila, etiqueta, formula, tipo, clave in RATIOS_06:
        r = str(fila)
        _f, _e, opt, lim, sentido, btipo = BM_POR_ETIQUETA[clave]
        bfila = str(_f)
        ws['B' + r] = etiqueta
        ws['B' + r]._style = copy.copy(est_label)
        motor.f(ws, 'C' + r, formula, fmts[tipo])
        # La referencia que se ENSEÑA sale de la misma tabla que el semáforo.
        motor.f(ws, 'D' + r, '=Benchmarks!$D$' + bfila)
        ws['D' + r].alignment = Alignment(horizontal='center')
        if sentido == 'menor':
            # ISNUMBER: si C trae el literal de aviso («Indica las ventas»),
            # Excel compararía TEXTO > NÚMERO como verdadero y encendería el
            # verde con el libro vacío (BLOQUEO 1-bis del crítico, 24-ago).
            estado = ('=IFERROR(IF(ISNUMBER($C' + r + '),IF($C' + r
                      + '<Benchmarks!$F$' + bfila
                      + ',"✅ Excelente",IF($C' + r + '<=Benchmarks!$G$'
                      + bfila + ',"⚠️ Aceptable","🔴 Alto")),"—"),"—")')
        else:
            estado = ('=IFERROR(IF(ISNUMBER($C' + r + '),IF($C' + r
                      + '>Benchmarks!$F$' + bfila
                      + ',"✅ Sano",IF($C' + r + '>=Benchmarks!$G$' + bfila
                      + ',"⚠️ Ajustado","🔴 Peligro")),"—"),"—")')
        motor.f(ws, 'E' + r, estado)
        motor.f(ws, 'F' + r, '=Benchmarks!$F$' + bfila, fmts[tipo])

    ws['B27'] = 'OTROS INDICADORES'
    ws['B27']._style = copy.copy(est_titulo)
    for fila, etiqueta, formula, tipo in OTROS_06:
        r = str(fila)
        ws['B' + r] = etiqueta
        ws['B' + r]._style = copy.copy(est_label)
        motor.f(ws, 'C' + r, formula, fmts[tipo])

    ws['B32'] = 'DATOS COMPLEMENTARIOS (existencias y desglose de gastos)'
    ws['B32']._style = copy.copy(est_titulo)
    for fila, etiqueta, valor, nota in COMPLEMENTARIOS_06:
        r = str(fila)
        ws['B' + r] = etiqueta
        ws['B' + r]._style = copy.copy(est_label)
        motor.aplicar_estilo(ws, 'C' + r, est_input, motor.FMT_EUR)
        motor.val(ws, 'C' + r, valor, motor.FMT_EUR, verde_=True)
        ws['D' + r] = nota
        ws['D' + r].font = Font(size=9, italic=True)
        ws['D' + r].alignment = Alignment(wrap_text=True, vertical='top')

    ws['B40'] = ('El GOP se mide ANTES del alquiler y de la amortización '
                 '(criterio USALI): por eso el alquiler tiene su propia línea '
                 'y el EBITDA % se calcula aparte.')
    ws['B40'].font = Font(size=9, italic=True)
    ws['B40'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A50'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A50'].font = Font(size=8)
    motor.anchos(ws, {'B': 40, 'C': 16, 'D': 30, 'E': 16, 'F': 14})
    informe.append('06!Ratios: RevPASH por plaza, food cost sobre ventas de '
                   'comida, consumo = Ei + compras − Ef, GOP definido y '
                   'semáforo en las nueve filas (DOM-07/DOM-18/DOM-22/RD-15)')


def _instrucciones_06(wb):
    ws = wb['Instrucciones']
    import re as _re
    for fila, clave in ((12, 'Food Cost %'), (13, 'Labor Cost %'),
                        (14, 'Prime Cost %'), (15, 'GOP %')):
        _f, etiqueta, opt, lim, sentido, tipo = BM_POR_ETIQUETA[clave]
        c, _d, e = _texto_benchmark(opt, lim, sentido, tipo)
        ws['B' + str(fila)] = ('▸ ' + etiqueta + ': óptimo ' + c
                               + ' · peligro ' + e
                               + ' (los números viven en Benchmarks!F:G).')
    motor.linea_instrucciones(
        ws, '▸ «Coste de comida — consumo» no son las facturas del mes: es '
            'existencias iniciales + compras − existencias finales. El mes '
            'que llenas la bodega, las compras se disparan y el consumo no.',
        rx=_re.compile(r'^▸ Los benchmarks del sector'))
    motor.linea_instrucciones(
        ws, '▸ El RevPASH se mide por PLAZA y hora (aforo × horas de '
            'servicio), no por m². «Horas de servicio» son aquellas en las '
            'que hay clientes sentados (p. ej. 6 h/día × 26 días = 156), no '
            'las que el local está abierto. Las ventas por m² tienen su '
            'propia fila.')


# ==========================================================================
# 07 · Informe de viabilidad
# ==========================================================================
PYL_07 = (
    (4, 'Facturación (sin IVA)', 'input'),
    (5, 'Coste de comida', 'input'),
    (6, 'Coste de bebida', 'input'),
    (7, 'Personal', 'input'),
    (8, 'Comisiones de plataformas (delivery)', 'input'),
    (9, 'Otros gastos operativos', 'input'),
    (10, 'TOTAL GASTOS', '=SUM({L}5:{L}9)'),
    (11, 'EBITDA', '={L}4-{L}10'),
    (12, 'Amortización', 'input'),
    (13, 'Intereses de la deuda', "='Financiación'!$D${F}"),
    (14, 'BAI (Beneficio Antes de Impuestos)', '={L}11-{L}12-{L}13'),
    (15, 'Impuesto sobre Sociedades', '=IF({L}14>0,{L}14*$C$29,0)'),
    (16, 'BENEFICIO NETO', '={L}14-{L}15'),
    # RX-06 — convención (a): TIR/VAN/payback DEL PROYECTO, sin apalancar.
    # El año 0 es la inversión TOTAL, así que los años 1-5 tienen que ser el
    # flujo libre del proyecto: EBITDA − IS calculado SIN intereses (que es
    # EBIT×(1−t)+amortización). Mezclar la inversión total con un flujo del
    # que ya se han restado los intereses —y en el que nunca sale la
    # devolución del principal— no medía ni el proyecto ni al inversor.
    (17, 'Flujo libre del proyecto (sin deuda)',
     '={L}11-MAX(0,({L}11-{L}12)*$C$29)'),
    (18, 'CASH FLOW OPERATIVO', '={L}16+{L}12'),
)
COLS_07 = ('B', 'C', 'D', 'E', 'F')
COLS_FLUJO = ('B', 'C', 'D', 'E', 'F', 'G')

# RX-04: el óptimo y el LÍMITE eran el mismo número en 6 de las 7 filas, así
# que el segundo IF repetía la condición del primero y la banda «⚠️» era
# código muerto: el informe anunciaba tres estados y sólo podía servir dos.
# La columna D sigue imprimiendo la referencia del ÓPTIMO (la que el analista
# espera leer); el límite es el suelo por debajo del cual ya no hay ámbar.
RATIOS_07 = (
    (4, 'Ratio de endeudamiento',
     '=IFERROR($C$16/($C$16+$C$15),"Indica deuda y fondos propios")',
     motor.FMT_PCT, 0.60, 0.70, 'menor', 'pct'),
    # V-01: el numerador era B18 (CASH FLOW OPERATIVO = beneficio neto +
    # amortización), que ya tiene los intereses restados, contra un
    # denominador que es el servicio COMPLETO de la deuda: los intereses se
    # descontaban DOS veces. Medido con EBITDA 36.000, amortización 15.000,
    # intereses 6.000 y servicio 16.103,59 → 1,63 en vez de ~1,9. No da verdes
    # falsos (sesga a la baja), pero sí ROJOS falsos en el documento que va al
    # banco. El numerador pasa a la fila 17, el FLUJO LIBRE DEL PROYECTO que
    # creó RX-06: EBITDA − IS sobre el EBIT, o sea EBIT×(1−t)+amortización.
    # Frente al otro candidato canonico (EBITDA − IS contable, B11−B15) se
    # elige la fila 17 por dos razones: es una fila IMPRESA que el analista
    # puede trazar en Proyecciones —B11−B15 no existe como línea— y su
    # impuesto no depende del apalancamiento, así que pedir más deuda no
    # infla el propio ratio que mide si puedes pagarla (con el ejemplo de
    # arriba: 30.750/16.103,59 = 1,91 frente a 32.250/16.103,59 = 2,00; la
    # fila 17 es la lectura conservadora, que es la que corresponde aquí).
    (5, 'Cobertura de deuda (DSCR)',
     "=IFERROR(Proyecciones!$B$17/'Financiación'!$C$12,\"Indica el préstamo\")",
     motor.FMT_X, 1.25, 1.15, 'mayor', 'x'),
    (6, 'Ratio de liquidez corriente',
     '=IFERROR($C$13/$C$14,"Indica el balance")', motor.FMT_X, 1.00, 0.90,
     'mayor', 'x'),
    (7, 'Margen EBITDA / Ventas',
     '=IFERROR(Proyecciones!$B$11/Proyecciones!$B$4,"Indica la facturación")',
     motor.FMT_PCT, 0.15, 0.10, 'mayor', 'pct'),
    (8, 'ROI (Return on Investment)',
     '=IFERROR(Proyecciones!$B$16/$C$17,"Indica la inversión")',
     motor.FMT_PCT, 0.10, 0.05, 'mayor', 'pct'),
    (9, 'Fondos propios / Inversión',
     '=IFERROR($C$15/$C$17,"Indica la inversión")', motor.FMT_PCT, 0.30,
     0.20, 'mayor', 'pct'),
    (10, 'Deuda / EBITDA',
     '=IFERROR($C$16/Proyecciones!$B$11,"Indica el EBITDA")', motor.FMT_X,
     3.5, 4.5, 'menor', 'x'),
)
BALANCE_07 = (
    (13, 'Activo corriente (€)', 'input'),
    (14, 'Pasivo corriente (€)', 'input'),
    (15, 'Fondos propios (€)', 'input'),
    (16, 'Deuda financiera viva (€)', 'input'),
    (17, 'Inversión total del proyecto (€)',
     "='Resumen Ejecutivo'!$C$13"),
)


def _financiacion_07(wb, informe):
    if 'Financiación' in wb.sheetnames:
        ws = wb['Financiación']
    else:
        ws = wb.create_sheet('Financiación', 3)
    motor.limpiar_rango(ws, 'A1:G30')
    ws['B1'] = 'Financiación — Cuadro de Amortización'
    ws['B1'].font = Font(bold=True, size=14)
    ws['B2'] = ('AI Chef Pro · aichef.pro — Kit Plan Financiero para '
                'Restaurantes')
    ws['B2'].font = Font(size=9)
    ws['B3'] = 'DATOS DEL PRÉSTAMO'
    ws['B3'].font = Font(bold=True)
    datos = (
        (4, 'Importe solicitado (€)', 0, motor.FMT_EUR,
         'El que pides al banco.'),
        (5, 'Tipo de interés nominal anual (%)', 0.06, motor.FMT_PCT,
         'TIN, no TAE. (valor orientativo — cámbialo)'),
        (6, 'Plazo (años)', 8, motor.FMT_ENT,
         'Plazo TOTAL, carencia incluida. (valor orientativo — cámbialo)'),
        (7, 'Carencia de capital (años)', 0, motor.FMT_ENT,
         'Años iniciales en los que sólo pagas intereses; después la cuota '
         'se calcula sobre el plazo restante.'),
    )
    for fila, etiqueta, valor, fmt, nota in datos:
        r = str(fila)
        ws['B' + r] = etiqueta
        motor.val(ws, 'C' + r, valor, fmt, verde_=True)
        ws['D' + r] = nota
        ws['D' + r].font = Font(size=9, italic=True)
    ws['B8'] = 'CUOTA ANUAL (€)'
    ws['B8'].font = Font(bold=True)
    # Anualidad ALGEBRAICA: `PAGO`/`PMT` está prohibido —pycel no lo implementa
    # y su IFERROR no lo atrapa, así que la celda se quedaría sin caché y en
    # blanco en Vista previa—. Verificado: 100.000 € al 5 % en 60 meses →
    # 1.887,12 €/mes.
    # RX-07: la cuota se calcula sobre el plazo que QUEDA tras la carencia
    # (C6 − C7). Con C7 = 0 la fórmula es idéntica a la anterior.
    # V-02: y con C7 ≥ C6 el exponente se vuelve positivo o nulo y la anualidad
    # deja de existir. Medido antes de este fix, préstamo de 100.000 € al 6 %
    # con C6=8: C7=10 imprimía una «CUOTA ANUAL» de −48.543,69 € y C7=8 una de
    # 0,00 € que amortizaba cero durante cinco años — las dos en la portada
    # financiera de un informe que se entrega a un banco. La DV de C7 es
    # «decimal ≥ 0» y no tiene techo, así que el input es alcanzable. Se anula
    # EN ORIGEN, como en 06!Ratios!C17 (RX-03): el mensaje sustituye al número
    # y protege a CUALQUIER consumidor, no sólo al que hoy se conoce.
    motor.f(ws, 'C8',
            '=IF(($C$6-$C$7)<=0,'
            '"La carencia no puede igualar ni superar el plazo",'
            'IFERROR($C$4*$C$5/(1-(1+$C$5)^(-($C$6-$C$7))),0))',
            motor.FMT_EUR)
    ws['C8'].font = Font(bold=True)
    ws['D8'] = ('Anualidad constante (sistema francés), calculada sobre el '
                'plazo posterior a la carencia. La carencia tiene que ser '
                'MENOR que el plazo: si no, no queda ningún año sobre el que '
                'repartir el capital y aquí sale un aviso en vez de una '
                'cifra.')
    ws['D8'].font = Font(size=9, italic=True)

    ws['B10'] = 'CUADRO DE AMORTIZACIÓN (años 1-5)'
    ws['B10'].font = Font(bold=True)
    cabecera = ('Año', 'Capital pendiente inicio (€)', 'Cuota anual (€)',
                'Intereses (€)', 'Amortización de capital (€)',
                'Capital pendiente fin (€)')
    for i, texto in enumerate(cabecera):
        cel = ws.cell(row=11, column=1 + i)
        cel.value = texto
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center', wrap_text=True)
    for i in range(5):
        r = str(12 + i)
        ws['A' + r] = i + 1
        if i == 0:
            motor.f(ws, 'B' + r, '=$C$4', motor.FMT_EUR)
        else:
            motor.f(ws, 'B' + r, '=$F$' + str(11 + i), motor.FMT_EUR)
        # RX-07: durante la carencia el SERVICIO del año son sólo intereses
        # (el capital pendiente no baja); la cuota francesa arranca el año
        # C7+1. Todo lo que lee la cuota —el DSCR de Ratios (C12) y la línea
        # de tesorería del 03— lee esta columna, no el C8 a secas.
        # V-02: con la carencia inválida C8 es TEXTO. El IF de la cuota NO
        # basta para absorberlo: en cuanto existe un año i > C7 (p. ej. C6=3
        # y C7=3, que deja los años 4 y 5 fuera de la carencia) la rama else
        # devuelve el texto y `E = C − D` daría #¡VALOR!. Así que la columna
        # de la cuota abstiene con «—» y la de amortización cae a 0 —que es
        # además lo correcto: si la carencia se come el plazo no se amortiza
        # nada—. Las columnas B, D y F siguen SIEMPRE numéricas a propósito:
        # de D salen los intereses del P&L (Proyecciones!B13) y de F el
        # capital pendiente que encadena el año siguiente; un «—» ahí
        # propagaría el #¡VALOR! al P&L entero.
        # Vencimiento (29-ago, orquestador): el cuadro pinta siempre 5 años,
        # así que un préstamo a 3 seguía amortizando en los años 4 y 5 y el
        # capital pendiente se volvía NEGATIVO. Pasado el plazo ($A > $C$6)
        # cuota, intereses y amortización caen a 0 (numéricos, no «—»).
        motor.f(ws, 'C' + r,
                '=IF($C$6<=$C$7,"—",IF($A' + r + '>$C$6,0,IF($A' + r
                + '<=$C$7,$B' + r + '*$C$5,$C$8)))',
                motor.FMT_EUR)
        motor.f(ws, 'D' + r, '=IF($A' + r + '>$C$6,0,$B' + r + '*$C$5)',
                motor.FMT_EUR)
        motor.f(ws, 'E' + r,
                '=IF($C$6<=$C$7,0,IF($A' + r + '>$C$6,0,$C' + r + '-$D' + r
                + '))', motor.FMT_EUR)
        motor.f(ws, 'F' + r, '=$B' + r + '-$E' + r, motor.FMT_EUR)
    ws['B18'] = ('La cuota de la columna C es la que se copia a '
                 '03!«Flujo Mensual», fila 23 (dividida entre 12), y la que '
                 'divide el DSCR de la pestaña Ratios.')
    ws['B18'].font = Font(size=9, italic=True)
    ws['B19'] = ('Los intereses de la columna D alimentan la fila «Intereses '
                 'de la deuda» de Proyecciones: no los teclees dos veces.')
    ws['B19'].font = Font(size=9, italic=True)
    ws['B20'] = ('Con carencia, los primeros años la columna «Amortización de '
                 'capital» sale a 0 y el capital pendiente no baja: es lo que '
                 'de verdad pasa en un ICO de apertura.')
    ws['B20'].font = Font(size=9, italic=True)
    ws['A22'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A22'].font = Font(size=8)
    motor.anchos(ws, {'A': 8, 'B': 30, 'C': 20, 'D': 18, 'E': 26, 'F': 24})
    informe.append("07: hoja 'Financiación' con anualidad algebraica, "
                   'carencia modelada y aviso en C8 cuando la carencia iguala '
                   'o supera el plazo; de ahí salen los intereses del P&L y '
                   'el DSCR (DOM-14/DOM-11/RX-07/V-02)')


def _proyecciones_07(wb, informe):
    ws = wb['Proyecciones']
    # B8 es TOTAL GASTOS en la v1.1 y «Comisiones de plataformas» (input) en
    # la v2.0; B10 es un valor calculado en las dos. El título se construye,
    # no se copia: A18 cambia de papel.
    est_label = copy.copy(ws['A4']._style)
    est_input = copy.copy(ws['B4']._style)
    est_total = copy.copy(ws['B10']._style)

    motor.limpiar_rango(ws, 'A4:G40')

    for fila, etiqueta, regla in PYL_07:
        r = str(fila)
        ws['A' + r] = etiqueta
        ws['A' + r]._style = copy.copy(est_label)
        for i, L in enumerate(COLS_07):
            coord = L + r
            if regla == 'input':
                motor.aplicar_estilo(ws, coord, est_input, motor.FMT_EUR)
                motor.val(ws, coord, 0, motor.FMT_EUR, verde_=True)
            else:
                motor.aplicar_estilo(ws, coord, est_total, motor.FMT_EUR)
                motor.f(ws, coord,
                        regla.replace('{L}', L).replace('{F}', str(12 + i)),
                        motor.FMT_EUR)
    ws['A12'].value = 'Amortización'
    ws['G12'] = ('Cópiala de 04!Resumen, columna «Dotación anual (€)» de la '
                 'fila INVERSIÓN TOTAL.')
    ws['G12'].font = Font(size=9, italic=True)
    ws['G17'] = ('EBITDA menos el Impuesto de Sociedades calculado SIN '
                 'intereses. Es la fila que alimenta la TIR, el VAN y el '
                 'payback de abajo.')
    ws['G17'].font = Font(size=9, italic=True)

    # --- bloque de flujos, ahora CON el año 0 -------------------------------
    ws['A20'] = 'FLUJOS DEL PROYECTO E INDICADORES DE RENTABILIDAD'
    ws['A20'].font = Font(bold=True, size=11)
    encabezados = ('Concepto', 'Año 0', 'Año 1', 'Año 2', 'Año 3', 'Año 4',
                   'Año 5')
    for i, texto in enumerate(encabezados):
        cel = ws.cell(row=21, column=1 + i)
        cel.value = texto
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center')

    ws['A22'] = 'Flujo de caja del proyecto'
    ws['A22']._style = copy.copy(est_label)
    motor.f(ws, 'B22', "=-'Resumen Ejecutivo'!$C$13", motor.FMT_EUR)
    for i, L in enumerate(COLS_FLUJO[1:]):
        # RX-06: la fila 17 (flujo libre del proyecto), no la 18 (cash flow
        # operativo, que ya lleva los intereses restados).
        motor.f(ws, L + '22', '=' + COLS_07[i] + '17', motor.FMT_EUR)
    ws['A23'] = 'Flujo acumulado'
    ws['A23']._style = copy.copy(est_label)
    motor.f(ws, 'B23', '=B22', motor.FMT_EUR)
    for i in range(1, 6):
        L, prev = COLS_FLUJO[i], COLS_FLUJO[i - 1]
        motor.f(ws, L + '23', '=$' + prev + '23+' + L + '22', motor.FMT_EUR)

    ws['A24'] = 'TIR del proyecto (sin apalancamiento)'
    ws['A24']._style = copy.copy(est_label)
    motor.f(ws, 'B24', '=IFERROR(IRR($B$22:$G$22),"—")', motor.FMT_PCT)
    ws['B24'].font = Font(bold=True)
    ws['A25'] = 'VAN del proyecto'
    ws['A25']._style = copy.copy(est_label)
    motor.f(ws, 'B25',
            '=IFERROR(NPV($C$27,$C$22:$G$22)+$B$22,"—")', motor.FMT_EUR)
    ws['B25'].font = Font(bold=True)
    ws['A26'] = 'Payback del proyecto (años)'
    ws['A26']._style = copy.copy(est_label)
    motor.f(ws, 'B26',
            '=IFERROR(IF($G$23<0,"No se recupera en 5 años",'
            'COUNTIF($B$23:$G$23,"<0")-1+(-INDEX($B$23:$G$23,'
            'COUNTIF($B$23:$G$23,"<0")))/INDEX($B$22:$G$22,'
            'COUNTIF($B$23:$G$23,"<0")+1)),"—")', '0.00" años"')
    ws['B26'].font = Font(bold=True)

    ws['A27'] = 'Tasa de descuento anual (%) — la usa el VAN'
    ws['A27']._style = copy.copy(est_label)
    motor.val(ws, 'C27', 0.08, motor.FMT_PCT, verde_=True)
    ws['A28'] = ('Es el coste de capital que exige quien pone el dinero. Un '
                 '8 % es la referencia habitual en hostelería; súbela si tu '
                 'riesgo es mayor.')
    ws['A28'].font = Font(size=9, italic=True)
    ws['A29'] = 'Tipo del Impuesto sobre Sociedades (%)'
    ws['A29']._style = copy.copy(est_label)
    motor.val(ws, 'C29', 0.25, motor.FMT_PCT, verde_=True)
    ws['A30'] = ('Tipo general 25 %. Entidad de NUEVA CREACIÓN: 15 % los dos '
                 'primeros ejercicios con base positiva (art. 29.1 LIS). Si '
                 'tributas por IRPF (autónomo), pon 0 aquí y calcula el IRPF '
                 'fuera.')
    ws['A30'].font = Font(size=9, italic=True)
    ws['A32'] = ('La inversión del año 0 sale del Resumen Ejecutivo: este '
                 'informe se entrega a un tercero y ningún dato de ejemplo '
                 'queda precargado con aspecto de dato real.')
    ws['A32'].font = Font(size=9, italic=True)
    ws['A33'] = ('TIR/VAN/payback se calculan sobre el flujo libre del '
                 'proyecto, sin deuda ni intereses: miden si el negocio se '
                 'sostiene por sí mismo. La capacidad de pagar el préstamo se '
                 'mide con el DSCR de la hoja Ratios.')
    ws['A33'].font = Font(size=9, italic=True)
    ws['A45'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A45'].font = Font(size=8)
    motor.anchos(ws, {'A': 38, 'G': 30})
    informe.append('07!Proyecciones: P&L con fórmulas, fila de flujo libre '
                   'del proyecto (sin deuda) que alimenta los flujos con año '
                   '0, TIR, VAN con tasa en celda, payback y tipo del IS en '
                   'celda (DOM-02/DOM-03/DOM-19/DOM-21/RX-06)')


def _ratios_07(wb, informe):
    ws = wb['Ratios']
    motor.limpiar_rango(ws, 'A3:H30')
    cabecera = ('Ratio', 'Valor', 'Referencia Bancaria', 'Estado',
                'Óptimo (nº)', 'Límite (nº)', 'Sentido')
    for i, texto in enumerate(cabecera):
        cel = ws.cell(row=3, column=2 + i)
        cel.value = texto
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center', wrap_text=True)
    for fila, etiqueta, formula, fmt, opt, lim, sentido, tipo in RATIOS_07:
        r = str(fila)
        ws['B' + r] = etiqueta
        motor.f(ws, 'C' + r, formula, fmt)
        if sentido == 'menor':
            ref = '< ' + (_pct(opt) if tipo == 'pct'
                          else str(opt) + 'x')
            # ISNUMBER: con el libro en blanco C es «Indica el préstamo» y
            # Excel da TEXTO >= NÚMERO como verdadero → cinco ✅ sin datos en
            # el informe que va al banco (BLOQUEO 1 del crítico, 24-ago).
            estado = ('=IFERROR(IF(ISNUMBER($C' + r + '),IF($C' + r + '<=$F$'
                      + r + ',"✅",IF($C' + r + '<=$G$' + r
                      + ',"⚠️","🔴")),"—"),"—")')
        else:
            ref = '> ' + (_pct(opt) if tipo == 'pct'
                          else str(opt) + 'x')
            estado = ('=IFERROR(IF(ISNUMBER($C' + r + '),IF($C' + r + '>=$F$'
                      + r + ',"✅",IF($C' + r + '>=$G$' + r
                      + ',"⚠️","🔴")),"—"),"—")')
        ws['D' + r] = ref
        ws['D' + r].alignment = Alignment(horizontal='center')
        motor.f(ws, 'E' + r, estado)
        motor.val(ws, 'F' + r, opt, fmt)
        motor.val(ws, 'G' + r, lim, fmt)
        ws['H' + r] = ('menor es mejor' if sentido == 'menor'
                       else 'mayor es mejor')
        ws['H' + r].font = Font(size=9, italic=True)

    ws['B12'] = 'DATOS DE BALANCE Y DEUDA'
    ws['B12'].font = Font(bold=True)
    for fila, etiqueta, regla in BALANCE_07:
        r = str(fila)
        ws['B' + r] = etiqueta
        if regla == 'input':
            motor.val(ws, 'C' + r, 0, motor.FMT_EUR, verde_=True)
        else:
            motor.f(ws, 'C' + r, regla, motor.FMT_EUR)

    ws['B20'] = ('El margen EBITDA usa los MISMOS umbrales que el dashboard '
                 '06 (óptimo > 15 %, ajustado 10-15 %): antes este informe '
                 'pedía «> 12 %» y el mismo restaurante aprobaba en un libro '
                 'y quedaba en ámbar en el otro.')
    ws['B20'].font = Font(size=9, italic=True)
    ws['B20'].alignment = Alignment(wrap_text=True, vertical='top')
    # V-01: la leyenda del DSCR. B21 estaba vacía y cae dentro del
    # `limpiar_rango(A3:H30)` de arriba, así que no desplaza nada y la 2.ª
    # pasada la reescribe igual.
    ws['B21'] = ('El DSCR divide el FLUJO LIBRE DEL PROYECTO del año 1 '
                 '(Proyecciones, fila 17: EBITDA menos el Impuesto sobre '
                 'Sociedades, antes de intereses) entre el servicio de deuda '
                 'de ese mismo año (Financiación, columna «Cuota anual», año '
                 '1). El numerador va ANTES de la deuda a propósito: si le '
                 'restaras los intereses, estarías descontándolos dos veces.')
    ws['B21'].font = Font(size=9, italic=True)
    ws['B21'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A24'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A24'].font = Font(size=8)
    motor.anchos(ws, {'B': 34, 'C': 16, 'D': 24, 'E': 12, 'F': 14, 'G': 14,
                      'H': 18})
    informe.append('07!Ratios: los siete se calculan, con el DSCR dividiendo '
                   'el flujo libre del proyecto (Proyecciones!B17, antes de '
                   'intereses) entre el servicio de deuda del año 1, y '
                   'columna Estado (DOM-11/TEC-19/RD-29/V-01)')


def _resumen_ejecutivo_07(wb, informe):
    ws = wb['Resumen Ejecutivo']
    derivadas = (
        ('C15', "='Financiación'!$C$4", motor.FMT_EUR),
        # RX-09: la portada del informe decía «0,00 € a 8,0 años». Sin
        # importe no hay plazo que enseñar.
        ('C16', "=IF('Financiación'!$C$4=0,\"—\",'Financiación'!$C$6)",
         '0.0" años"'),
        ('C17', '=Proyecciones!$B$4', motor.FMT_EUR),
        ('C18', '=Proyecciones!$B$11', motor.FMT_EUR),
        ('C19', '=Proyecciones!$B$26', '0.00" años"'),
    )
    for coord, formula, fmt in derivadas:
        motor.f(ws, coord, formula, fmt)
        ws[coord].fill = motor.PatternFill()      # deja de ser input
    ws['B29'] = ('Financiación, plazo, facturación, EBITDA y payback se '
                 'calculan solos desde Financiación y Proyecciones: no se '
                 'teclean dos veces.')
    ws['B29'].font = Font(size=9, italic=True)
    informe.append('07!Resumen Ejecutivo: cinco líneas dejan de ser input y '
                   'se calculan (TEC-20)')


# ==========================================================================
# BONUS-08 · Simulador
# ==========================================================================
ESCENARIOS_08 = {
    'B': {'ticket': 18, 'cubiertos': 40, 'dias': 25, 'food': 0.40,
          'labor': 0.32, 'fijos': 6500},
    'C': {'ticket': EJ['ticket'], 'cubiertos': EJ['cubiertos_dia'],
          'dias': EJ['dias_apertura'], 'food': 0.35,
          'labor': EJ['labor_cost_pct'],
          'fijos': EJ['fijos_sin_personal']},
    'D': {'ticket': 28, 'cubiertos': 75, 'dias': 27, 'food': 0.30,
          'labor': 0.25, 'fijos': 5900},
}


def _bonus_08(wb, informe):
    ws = wb['Simulador']
    # La variable de la fila 9 no es «food cost» a secas: incluye las
    # comisiones de delivery, igual que el `% Coste variable` del 02. Con dos
    # nombres para la misma magnitud, los dos libros daban EBITDA distinto
    # para el mismo restaurante.
    ws['A9'] = '% Coste variable (food + comisiones de delivery)'
    ws['A11'] = ('Costes fijos mensuales SIN personal y SIN cuota de préstamo '
                 '(alquiler, suministros, seguros, marketing, gestoría, '
                 'otros)')
    ws['A12'] = ('Si copias los costes fijos del 02, RESTA antes las nóminas y '
                 'la cuota: el personal ya entra arriba como % sobre ventas y '
                 'el EBITDA se mide antes del servicio de deuda.')
    ws['A12'].font = Font(size=9, italic=True)
    fmts = {'ticket': motor.FMT_EUR, 'cubiertos': motor.FMT_ENT,
            'dias': motor.FMT_ENT, 'food': motor.FMT_PCT,
            'labor': motor.FMT_PCT, 'fijos': motor.FMT_EUR}
    filas = {'ticket': 6, 'cubiertos': 7, 'dias': 8, 'food': 9, 'labor': 10,
             'fijos': 11}
    for col, valores in ESCENARIOS_08.items():
        for clave, valor in valores.items():
            motor.val(ws, col + str(filas[clave]), valor, fmts[clave],
                      verde_=True)
    # V-04: las tres últimas celdas del patrón `=IF(x=0,0,…)` del kit. Este
    # libro SÍ se entrega con datos de ejemplo, así que hoy no imprime ningún
    # «0,0 %» falso; el guardián sólo salta si el cliente pone a cero el ticket
    # o los cubiertos de un escenario, y entonces conviene que el margen se
    # calle en vez de afirmar un 0,0 %. Consumidor censado: `Comparativa`!C8:E8
    # (`=Simulador!B20`, un pase directo sin aritmética) y nada más — el
    # gráfico de la Comparativa cuelga de C7:E7 (EBITDA anual) y el formato
    # condicional, de C7:E7 y de Simulador!B19:D19; la fila 20 no alimenta
    # ninguno de los dos. Un texto vacío se propaga como texto vacío.
    for L in ('B', 'C', 'D'):
        motor.f(ws, L + '20', '=IF(' + L + '14=0,"",' + L + '19/' + L + '14)',
                motor.FMT_PCT)
    ws['A22'] = ('Base común con el 02: mismo ticket (22 €), mismos '
                 'cubiertos/día (55), mismos días (26), mismo % de coste '
                 'variable (35 %) y los mismos costes fijos sin personal '
                 '(6.200 €). El EBITDA base coincide: ~17 % sobre ventas en '
                 'los dos libros.')
    ws['A22'].font = Font(size=9, italic=True)
    motor.anchos(ws, {'A': 56})
    informe.append('BONUS-08: fin de la doble contabilización del personal, '
                   'escenarios homogéneos con el 02 y margen EBITDA que se '
                   'calla sin facturación (DOM-06/TEC-21/TEC-22/V-04)')


# ==========================================================================
# BONUS-09 · Checklist
# ==========================================================================
FASE_7 = (
    'Inscribir la empresa en la Seguridad Social y obtener el CCC',
    'Alta del promotor/a en el RETA (o régimen que corresponda)',
    'Comunicación de apertura del centro de trabajo a la autoridad laboral',
    'Contratos según el convenio provincial de hostelería',
    'Alta de los trabajadores en la SS ANTES del inicio de la jornada',
    'Registro de jornada y calendario de nóminas y del modelo 111',
)
# RD-07: la banda de la fase 3 usaba el MISMO E8F5E9 que los editables, así
# que se desbloqueaba por accidente mientras la columna Estado —la única que
# el usuario tiene que tocar— quedaba cerrada. Colores nuevos que no colisionan.
COLOR_FASE = {
    '1. Constitución': 'E3F2FD',
    '2. Financiación': 'FFF3E0',
    '3. Licencias': 'DCEDC8',
    '4. Proveedores': 'F3E5F5',
    '5. Seguros': 'FFEBEE',
    '6. Tesorería': 'E0F7FA',
    '7. Personal y obligaciones laborales': 'FFFDE7',
}
TAREAS_TOTAL = 54
FILA_ULTIMA_TAREA = 4 + TAREAS_TOTAL          # 58
FILA_RESUMEN = 66


def _bonus_09(wb, informe):
    ws = wb['Checklist']
    est_fase = copy.copy(ws['B5']._style)
    est_tarea = copy.copy(ws['C5']._style)
    est_estado = copy.copy(ws['F5']._style)

    ws['A1'] = 'Checklist Financiero Pre-Apertura — 54 Tareas'

    # Se limpia desde la primera fila de la fase 7 hasta el final: ahí vivía el
    # bloque RESUMEN de la v1.1 (filas 54-59) y el © de la 61.
    motor.limpiar_rango(ws, 'A53:G' + str(FILA_RESUMEN + 12))

    # Fase 7 (DOM-20): la checklist financiera no tenía NI UNA tarea de alta
    # laboral —todas con sanción directa e impacto de caja el primer mes—
    # mientras la fase 6 detallaba la «política de propinas».
    fase7 = '7. Personal y obligaciones laborales'
    for i, tarea in enumerate(FASE_7):
        r = 53 + i
        ws.cell(row=r, column=1, value=49 + i)
        cel = ws.cell(row=r, column=2, value=fase7)
        cel._style = copy.copy(est_fase)
        tar = ws.cell(row=r, column=3, value=tarea)
        tar._style = copy.copy(est_tarea)
        est = ws.cell(row=r, column=6, value='Pendiente')
        est._style = copy.copy(est_estado)

    # RD-07: la banda de la fase 3 usaba el MISMO E8F5E9 que los editables, así
    # que la protección la abría por accidente mientras cerraba la columna
    # Estado, que es la única que el usuario tiene que tocar.
    for r in range(5, FILA_ULTIMA_TAREA + 1):
        cel = ws.cell(row=r, column=2)
        color = COLOR_FASE.get(cel.value)
        if color:
            cel.fill = motor.PatternFill('solid', fgColor=color)

    # Las cuatro columnas que el usuario rellena, editables de verdad, con seis
    # filas libres (59-64) para las tareas que añada él.
    motor.marcar_editable(ws, 'D5:G64')
    for r in range(5, 65):
        ws.cell(row=r, column=5).number_format = motor.FMT_FECHA

    # DV de estado: hasta la 64 (no la 52) y con showErrorMessage=True. La de
    # la v1.1 dejaba pasar «completado» o «COMPLETADA » —que el formato
    # condicional pintaba de verde— mientras el contador seguía en 0 %.
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if dv.type != 'list']
    dv = DataValidation(type='list',
                        formula1='"Pendiente,En curso,Completada"',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle='Estado no válido',
                        error='Elige Pendiente, En curso o Completada.')
    ws.add_data_validation(dv)
    dv.add('F5:F64')

    # Bloque de resumen FUERA del rango A5:A64 que cuenta las tareas.
    ws['B' + str(FILA_RESUMEN)] = 'RESUMEN'
    ws['B' + str(FILA_RESUMEN)].font = Font(bold=True)
    # COUNTA está prohibida (pycel no la implementa): COUNTIF(rango,"<>") da el
    # mismo recuento y sí se evalúa.
    filas = (
        (FILA_RESUMEN + 1, 'Total tareas:', '=COUNTIF($A$5:$A$64,"<>")',
         motor.FMT_ENT),
        (FILA_RESUMEN + 2, 'Completadas:',
         '=COUNTIF($F$5:$F$64,"Completada")', motor.FMT_ENT),
        (FILA_RESUMEN + 3, 'En curso:', '=COUNTIF($F$5:$F$64,"En curso")',
         motor.FMT_ENT),
        (FILA_RESUMEN + 4, 'Pendientes:', '=COUNTIF($F$5:$F$64,"Pendiente")',
         motor.FMT_ENT),
        (FILA_RESUMEN + 5, '% Completado:',
         '=IFERROR(COUNTIF($F$5:$F$64,"Completada")/$C$'
         + str(FILA_RESUMEN + 1) + ',0)', motor.FMT_PCT),
    )
    for fila, etiqueta, formula, fmt in filas:
        ws['B' + str(fila)] = etiqueta
        motor.f(ws, 'C' + str(fila), formula, fmt)
    ws['B' + str(FILA_RESUMEN + 7)] = (
        'El total se CUENTA, no se escribe: si añades tareas tuyas en las '
        'filas libres (hasta la 64), el porcentaje sigue siendo cierto.')
    ws['B' + str(FILA_RESUMEN + 7)].font = Font(size=9, italic=True)
    ws['A' + str(FILA_RESUMEN + 9)] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A' + str(FILA_RESUMEN + 9)].font = Font(size=8)
    informe.append('BONUS-09: fase 7 laboral (54 tareas), contador honesto, '
                   'DV bloqueante hasta la fila 64 y columna Estado '
                   'desbloqueada (DOM-20/DOM-26/TEC-24/RD-07)')


def _instrucciones_09(wb):
    import re as _re
    ws = wb['Instrucciones']
    ws['B2'] = 'BONUS 09 · Checklist Financiero Pre-Apertura — 54 Tareas'
    motor.linea_instrucciones(
        ws, '▸ 54 tareas organizadas en 7 fases: Constitución, Financiación, '
            'Licencias, Proveedores, Seguros, Tesorería y Personal.',
        rx=_re.compile(r'^▸ 48 tareas'))
    motor.linea_instrucciones(
        ws, '▸ Fase 7: Personal y obligaciones laborales (SS, RETA, apertura '
            'del centro de trabajo, contratos del convenio, altas previas, '
            'registro de jornada y modelo 111).')
    motor.linea_instrucciones(
        ws, '▸ El «% Completado» cuenta las tareas que hay, no un número '
            'escrito a mano: si añades las tuyas hasta la fila 64, el '
            'porcentaje las incluye.')


# ==========================================================================
# API del grupo
# ==========================================================================
def post(wb, fname, cambios, registro):
    if fname.startswith('06'):
        _benchmarks_06(wb, cambios)
        _ratios_06(wb, cambios)
        _instrucciones_06(wb)
    elif fname.startswith('07'):
        _financiacion_07(wb, cambios)
        _proyecciones_07(wb, cambios)
        _ratios_07(wb, cambios)
        _resumen_ejecutivo_07(wb, cambios)
        import re as _re
        ws = wb['Instrucciones']
        motor.linea_instrucciones(
            ws, '▸ TIR, VAN y Payback se calculan solos en Proyecciones sobre '
                'el FLUJO LIBRE DEL PROYECTO (sin deuda ni intereses) y la '
                'inversión del Resumen Ejecutivo; la capacidad de pagar el '
                'préstamo la mide el DSCR de Ratios. La tasa de descuento es '
                'tuya (celda C27).',
            rx=_re.compile(r'^▸ TIR, VAN'))
        # RX-10: la lista de pestañas enumeraba CUATRO y se dejaba fuera
        # 'Financiación', que es la hoja nueva de la v2.0 — se la nombraba
        # cuatro líneas más abajo, tras un hueco. Se escribe en su orden real
        # de aparición: Resumen Ejecutivo · Proyecciones · Financiación ·
        # Ratios · Garantías.
        est_lista = copy.copy(ws['B14']._style)
        for fila, texto in (
                (15, '▸ Financiación: cuadro de amortización del préstamo. De '
                     'ahí salen los intereses del P&L, la cuota del DSCR y la '
                     'línea de cuota de la tesorería del 03.'),
                (16, '▸ Ratios: indicadores de solvencia y rentabilidad. El '
                     'DSCR compara el flujo libre del proyecto del año 1 '
                     '(Proyecciones, fila 17) con el servicio de deuda de ese '
                     'año (Financiación): mide si el negocio paga la cuota.'),
                (17, '▸ Garantías: avales y garantías ofrecidas.')):
            ws['B' + str(fila)] = texto
            ws['B' + str(fila)]._style = copy.copy(est_lista)
        motor.linea_instrucciones(
            ws, '▸ La amortización del inmovilizado se copia de 04!Resumen, '
                'columna «Dotación anual (€)».')
    elif fname.startswith('BONUS-08'):
        _bonus_08(wb, cambios)
    elif fname.startswith('BONUS-09'):
        _bonus_09(wb, cambios)
        _instrucciones_09(wb)
    return cambios


def demos(carpeta, origen, destino):
    """El 07 se entrega SIN datos de ejemplo (§4), así que la prueba de que
    calcula se hace sobre una copia con datos inyectados por pycel."""
    import contextlib
    import os
    import shutil

    import openpyxl
    from pycel import ExcelCompiler

    fuera = {'fallos': [], 'grupo_c': {}}

    def _ev(xl, ref):
        with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
            try:
                return xl.evaluate(ref)
            except Exception as e:                           # noqa: BLE001
                return 'ERR:' + type(e).__name__

    os.makedirs(destino, exist_ok=True)

    # --- 07: P&L, flujos, VAN, payback y DSCR con un proyecto real ----------
    f07 = '07-informe-viabilidad-bancos.xlsx'
    p07 = os.path.join(carpeta, f07)
    if os.path.isfile(p07):
        copia = os.path.join(destino, '_gc_' + f07)
        shutil.copy2(p07, copia)
        xl = ExcelCompiler(filename=copia)
        for ref in ("'Proyecciones'!B10", "'Proyecciones'!B11",
                    "'Proyecciones'!B18", "'Proyecciones'!B22",
                    "'Proyecciones'!G23", "'Proyecciones'!B25",
                    "'Proyecciones'!B26", "'Financiación'!C8",
                    "'Financiación'!D12", "'Ratios'!C5", "'Ratios'!C7",
                    "'Ratios'!E5"):
            _ev(xl, ref)
        xl.set_value("'Resumen Ejecutivo'!C13", 150000)
        xl.set_value("'Financiación'!C4", 100000)
        ventas = (300000.0, 330000.0, 360000.0, 390000.0, 420000.0)
        for i, L in enumerate(('B', 'C', 'D', 'E', 'F')):
            xl.set_value("'Proyecciones'!" + L + '4', ventas[i])
            xl.set_value("'Proyecciones'!" + L + '5', ventas[i] * 0.26)
            xl.set_value("'Proyecciones'!" + L + '6', ventas[i] * 0.05)
            xl.set_value("'Proyecciones'!" + L + '7', ventas[i] * 0.30)
            xl.set_value("'Proyecciones'!" + L + '8', ventas[i] * 0.03)
            xl.set_value("'Proyecciones'!" + L + '9', ventas[i] * 0.22)
            xl.set_value("'Proyecciones'!" + L + '12', 15000)
        cuota = _ev(xl, "'Financiación'!C8")
        ebitda1 = _ev(xl, "'Proyecciones'!B11")
        cf1 = _ev(xl, "'Proyecciones'!B18")
        flujo0 = _ev(xl, "'Proyecciones'!B22")
        acum5 = _ev(xl, "'Proyecciones'!G23")
        van = _ev(xl, "'Proyecciones'!B25")
        pb = _ev(xl, "'Proyecciones'!B26")
        dscr = _ev(xl, "'Ratios'!C5")
        estado_dscr = _ev(xl, "'Ratios'!E5")
        flujos = [flujo0] + [_ev(xl, "'Proyecciones'!" + L + '22')
                             for L in ('C', 'D', 'E', 'F', 'G')]
        tir = None
        if all(_es_num(x) for x in flujos):
            tir = motor.tir_newton(flujos)
        fuera['grupo_c']['07'] = {
            'cuota_anual_100000_al_6pct_8anios': cuota,
            'EBITDA_anio_1': ebitda1, 'cash_flow_operativo_anio_1': cf1,
            'flujo_anio_0': flujo0, 'flujo_acumulado_anio_5': acum5,
            'VAN_al_8pct': van, 'payback': pb, 'DSCR': dscr,
            'estado_DSCR': estado_dscr,
            'flujos': flujos,
            'TIR_newton_sobre_esos_flujos': tir,
        }
        if not (_es_num(ebitda1) and ebitda1 > 0):
            fuera['fallos'].append(
                '07: el EBITDA del año 1 no calcula (' + str(ebitda1) + ')')
        if flujo0 != -150000:
            fuera['fallos'].append(
                "07: el flujo del año 0 no sale del Resumen Ejecutivo ("
                + str(flujo0) + ')')
        if not _es_num(van):
            fuera['fallos'].append('07: el VAN no evalúa (' + str(van) + ')')
        if not _es_num(pb):
            fuera['fallos'].append(
                '07: el payback no evalúa (' + str(pb) + ')')
        if not (_es_num(dscr) and dscr > 0):
            fuera['fallos'].append('07: el DSCR no calcula (' + str(dscr)
                                   + ')')
        if tir is None:
            fuera['fallos'].append(
                '07: con flujos reales la TIR sigue sin existir')

    # --- 06: el umbral único y el RevPASH por plaza -------------------------
    f06 = '06-dashboard-ratios-financieros.xlsx'
    p06 = os.path.join(carpeta, f06)
    if os.path.isfile(p06):
        copia = os.path.join(destino, '_gc_' + f06)
        shutil.copy2(p06, copia)
        xl = ExcelCompiler(filename=copia)
        food = _ev(xl, "'Ratios'!C17")
        estado_food = _ev(xl, "'Ratios'!E17")
        revpash = _ev(xl, "'Ratios'!C22")
        gop = _ev(xl, "'Ratios'!C20")
        consumo = _ev(xl, "'Ratios'!C8")
        cociente = _ev(xl, "'Ratios'!C24")
        estado_cociente = _ev(xl, "'Ratios'!E24")
        wb = openpyxl.load_workbook(copia)
        ref_texto = wb['Ratios']['D17'].value
        fuera['grupo_c']['06'] = {
            'food_cost_sobre_ventas_de_comida': food,
            'estado_food_cost': estado_food,
            'referencia_que_se_enseña': ref_texto,
            'RevPASH_por_plaza_y_hora': revpash, 'GOP': gop,
            'consumo_de_comida': consumo,
            'coste_cubierto_entre_ticket': cociente,
            'estado_de_esa_fila': estado_cociente,
        }
        if not isinstance(estado_cociente, str) or estado_cociente == '':
            fuera['fallos'].append(
                '06: la fila «Coste por cubierto / Ticket» sigue sin semáforo '
                '(DOM-22)')
        if _es_num(revpash) and _es_num(food):
            # 31.460 € / (60 plazas × 360 h) = 1,456 €/plaza/hora.
            if abs(revpash - (VENTAS_EJ / (EJ['aforo_plazas']
                                           * EJ['horas_servicio_mes']))) > 0.01:
                fuera['fallos'].append(
                    '06: el RevPASH no divide por plazas × horas ('
                    + str(revpash) + ')')

    # --- BONUS-09: contador honesto -----------------------------------------
    f09 = 'BONUS-09-checklist-pre-apertura.xlsx'
    p09 = os.path.join(carpeta, f09)
    if os.path.isfile(p09):
        copia = os.path.join(destino, '_gc_' + f09)
        shutil.copy2(p09, copia)
        xl = ExcelCompiler(filename=copia)
        total = _ev(xl, "'Checklist'!C" + str(FILA_RESUMEN + 1))
        pct = _ev(xl, "'Checklist'!C" + str(FILA_RESUMEN + 5))
        _ev(xl, "'Checklist'!C" + str(FILA_RESUMEN + 2))
        xl.set_value("'Checklist'!F5", 'Completada')
        pct2 = _ev(xl, "'Checklist'!C" + str(FILA_RESUMEN + 5))
        wb = openpyxl.load_workbook(copia)
        wsc = wb['Checklist']
        estado_libre = sum(1 for r in range(5, 65)
                           if not wsc.cell(row=r, column=6).protection.locked)
        dvs = [dv for dv in wsc.data_validations.dataValidation
               if dv.type == 'list']
        fuera['grupo_c']['BONUS-09'] = {
            'total_contado': total, 'pct_inicial': pct,
            'pct_con_una_completada': pct2,
            'celdas_de_estado_desbloqueadas': estado_libre,
            'DV_lista': [str(d.sqref) + ' err=' + str(d.showErrorMessage)
                         for d in dvs],
        }
        if total != TAREAS_TOTAL:
            fuera['fallos'].append(
                'BONUS-09: el contador dice ' + str(total) + ' y las tareas '
                'son ' + str(TAREAS_TOTAL))
        if estado_libre != 60:
            fuera['fallos'].append(
                'BONUS-09: sólo ' + str(estado_libre) + '/60 celdas de Estado '
                'son editables (RD-07)')
        if not dvs or not all(d.showErrorMessage for d in dvs):
            fuera['fallos'].append(
                'BONUS-09: la DV de estado no es bloqueante')

    # --- BONUS-08: el escenario base deja de estar en pérdidas --------------
    f08 = 'BONUS-08-simulador-escenarios.xlsx'
    p08 = os.path.join(carpeta, f08)
    if os.path.isfile(p08):
        copia = os.path.join(destino, '_gc_' + f08)
        shutil.copy2(p08, copia)
        xl = ExcelCompiler(filename=copia)
        base_mes = _ev(xl, "'Simulador'!C19")
        base_anual = _ev(xl, "'Simulador'!C21")
        margen = _ev(xl, "'Simulador'!C20")
        opt = _ev(xl, "'Simulador'!D19")
        pes = _ev(xl, "'Simulador'!B19")
        fuera['grupo_c']['BONUS-08'] = {
            'EBITDA_base_mensual': base_mes, 'EBITDA_base_anual': base_anual,
            'margen_base': margen, 'EBITDA_optimista': opt,
            'EBITDA_pesimista_(negativo_a_proposito)': pes,
        }
        if not (_es_num(base_mes) and base_mes > 0):
            fuera['fallos'].append(
                'BONUS-08: el escenario BASE sigue en pérdidas ('
                + str(base_mes) + ')')
        if not (_es_num(margen) and 0.05 <= margen <= 0.20):
            fuera['fallos'].append(
                'BONUS-08: el margen del escenario base es ' + str(margen)
                + ', fuera del 5-20 % razonable')
    return fuera
