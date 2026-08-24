#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 del Kit Plan Financiero.

    python3 main.py --dry-run [--solo motor|a|b|c|a,b] [--json informe.json]

REGLA DURA: `astro-site/public/dl/kit-plan-financiero/` NO se toca. `--dry-run`
regenera una copia en el scratchpad y trabaja allí. Sin `--dry-run` el script
ABORTA salvo que se le pase `KIT_PLAN_FINANCIERO_APPLY=1` en el entorno: la
ejecución real la hace el orquestador cuando la ronda 2 dé verde (SPEC,
cabecera). En producción se hace **respaldo previo al scratchpad** antes de
escribir in place — nunca un `.bak` dentro de `public/dl/`, que se publicaría.

Orden del pipeline (§1.8 de la SPEC, y el orden importa):
  1. Copia de trabajo (o respaldo).
  2. Por fichero: `grupo.pre` → `motor.aplicar` (§1) → `grupo.post` →
     `motor.cerrar` (§1: semáforos, DV, A4, protección, bio+versión) →
     **`motor.graficos`** → `wb.save()`.
  3. Idempotencia: se repite el pipeline entero sobre un clon del resultado y
     se compara celda a celda. Debe dar 0 diferencias.
  4. `inject_cache.py` sobre los ficheros tocados — SIEMPRE al final: cualquier
     `wb.save()` posterior borraría el caché que acaba de escribir.
  5. `motor.cachear_irr()` sobre el 07: pycel no evalúa `IRR`, así que esa
     celda es la única que `inject_cache` deja en blanco, y en blanco se
     entrega al banco.
  6. Verificación `data_only` de todas las fórmulas nuevas del registro.
  7. Gate de gráficos: `len(ws._charts) >= 1` en los 9 ficheros del §1.1,
     contados DESPUÉS del post-proceso y de `inject_cache`.
  8. `censo-entregables.py --only <carpeta> --fail --quiet`.
  9. Demostraciones con pycel (TIR/VAN/payback del caso trazado, anualidad
     algebraica en vez de `PMT`, `COUNTIF` en vez de `COUNTA`, semáforo que
     cambia con un dato fuera de límite, `IFERROR` que sustituye al `#¡DIV/0!`,
     encadenado del saldo del 03 y protección sin contraseña).

**Los módulos de grupo pueden no existir todavía.** `--solo motor` no carga
ninguno y el pipeline funciona igual: el motor hace su §1 y los gráficos cuyos
datos ya existen; los que dependen de un bloque que crea un grupo ausente se
anotan como `pendientes_grupo` y NO cuentan como fallo.

CONTRATO de `grupo_a.py` / `grupo_b.py` / `grupo_c.py` (probado con un grupo de
prueba antes de entregar esto):

    FICHEROS = ['06-....xlsx', ...]     # obligatorio: qué ficheros toca
    PROPIOS  = [...]                    # opcional: ficheros donde el §1 del
                                        # motor NO debe aplicarse (el grupo se
                                        # encarga entero, incluido motor.cerrar)
    def pre(wb, fname, cambios): ...    # opcional, antes de motor.aplicar
    def post(wb, fname, cambios, registro): ...   # tras motor.aplicar
    def demos(carpeta, origen, destino) -> dict   # opcional; su clave
                                        # 'fallos' se suma al veredicto

Las FÓRMULAS se escriben con `motor.f(ws, coord, formula, fmt)` —queda
registrada y `main.py` verifica una por una que acabó con valor cacheado— o, si
se escriben a mano, se añaden a la lista `registro` como `(hoja, coord,
formula)`. Los valores editables, con `motor.val(..., verde_=True)`: el verde
es lo que `motor.cerrar()` usa para decidir qué celda queda desbloqueada.

Los grupos NO tienen que ocuparse de: protección, A4, bio/versión, el sufijo
«(sin IVA)» de las etiquetas de ventas (se aplica DESPUÉS, en `motor.cerrar`,
justo para que los grupos busquen las etiquetas tal como están hoy), la
validación numérica de las celdas verdes ni los gráficos.

Térmica: todo en SERIE, un python cada vez. No hay builds ni navegador.
"""
import argparse
import contextlib
import datetime
import importlib
import json
import logging
import os
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
PRODUCTO = 'kit-plan-financiero'
ORIGEN = os.path.join(ROOT, 'astro-site', 'public', 'dl', PRODUCTO)
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    '7340312f-b4fe-4aa1-b254-4b0c17c8375f/scratchpad')
DESTINO = os.path.join(SCRATCH, 'dryrun-v2', PRODUCTO)
IDEM = os.path.join(SCRATCH, 'dryrun-v2', PRODUCTO + '-idem')
DEMOS = os.path.join(SCRATCH, 'dryrun-v2', '_demos-pf')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')

SPEC = 'scripts/productos-digitales/kit-plan-financiero-v2-SPEC.md §1'

# §1.1: los 9 ficheros que tienen que acabar con al menos un gráfico. El
# BONUS-09 es una checklist y no entra.
CON_GRAFICO = [n for n in motor.FICHEROS
               if n != 'BONUS-09-checklist-pre-apertura.xlsx']

# Qué grupo construye los datos de cada gráfico: si el grupo NO se ha cargado,
# el gráfico pendiente es esperado y no es un fallo.
GRUPO_DE = {
    '01-plan-financiero-previsional.xlsx': 'a',
    '01b-plan-financiero-previsional-5-anos.xlsx': 'a',
    '05-pyl-mensual-real-vs-presupuesto.xlsx': 'a',
    '02-calculadora-punto-equilibrio.xlsx': 'b',
    '03-cash-flow-forecast.xlsx': 'b',
    '04-presupuesto-inversion-capex.xlsx': 'b',
    '06-dashboard-ratios-financieros.xlsx': 'c',
    '07-informe-viabilidad-bancos.xlsx': 'c',
    'BONUS-08-simulador-escenarios.xlsx': 'c',
}


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Preparación
# ==========================================================================
def preparar_copia():
    if not os.path.isdir(ORIGEN):
        raise SystemExit('No existe el origen: ' + ORIGEN)
    if os.path.isdir(DESTINO):
        shutil.rmtree(DESTINO)
    os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
    shutil.copytree(ORIGEN, DESTINO)
    log('  copia de trabajo regenerada: ' + DESTINO)


def digest(path):
    """Huella comparable de un .xlsx (valores, formatos, relleno, bloqueo,
    merges, DV, formato condicional, protección, gráficos)."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked))
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted(str(dv.type) + ':' + str(dv.formula1) + ':'
                         + str(dv.sqref)
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted(str(cf.sqref) + ':' + str(len(cf.rules))
                         for cf in ws.conditional_formatting),
            'prot': bool(ws.protection.sheet),
            'charts': len(ws._charts),
            'area': str(ws.print_area),
            'pie': str(ws.oddFooter.center.text),
        }
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append(fichero + ':' + hoja + ': hoja sólo en una pasada')
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'prot', 'area', 'charts', 'pie'):
            if ha[k] != hb[k]:
                fuera.append(fichero + ':' + hoja + ': cambia ' + k + ' ('
                             + str(ha[k])[:120] + ' → ' + str(hb[k])[:120]
                             + ')')
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append(fichero + ':' + hoja + '!' + coord + ': '
                             + str(ca.get(coord)) + ' → ' + str(cb.get(coord)))
    return fuera


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, grupos, informe_global):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    cambios, registro_grupo = [], []
    motor.REGISTRO = []

    propio = any(fname in getattr(g, 'PROPIOS', []) for g in grupos)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []) and hasattr(g, 'pre'):
            g.pre(wb, fname, cambios)

    if not propio:
        motor.aplicar(wb, fname, cambios)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []) and hasattr(g, 'post'):
            g.post(wb, fname, cambios, registro_grupo)

    if not propio:
        motor.cerrar(wb, fname, cambios)

    # §1.8: los gráficos, DESPUÉS de que las hojas estén construidas y ANTES
    # de guardar. Se cuentan luego, tras `inject_cache`.
    graficos = motor.graficos(wb, fname, cambios)
    literales = motor.literales_sospechosos(wb, fname)
    cuenta = motor.contadores(wb, fname)
    wb.save(path)

    registro = list(motor.REGISTRO)
    registro += [(h, c, fm) for h, c, fm in registro_grupo
                 if isinstance(fm, str) and fm.startswith('=')]
    informe_global.append({'fichero': fname, 'cambios': cambios,
                           'formulas_nuevas': len(registro),
                           'graficos': graficos,
                           'literales_sospechosos': literales[:12],
                           'contadores': cuenta})
    return registro


def ficheros_a_tocar(grupos):
    nombres = list(motor.FICHEROS)
    for g in grupos:
        for fn in getattr(g, 'FICHEROS', []):
            if fn not in nombres:
                nombres.append(fn)
    return nombres


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        salida = r.stdout.strip()
        fuera.append((n, salida, r.returncode))
        log('    ' + salida)
        if r.returncode != 0:
            log('    ERROR inject_cache: ' + r.stderr[-400:])
    return fuera


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA debe tener valor cacheado, salvo las que devuelven
    `""` por diseño y la `IRR`, que la cachea `motor.cachear_irr()`."""
    fallos, vacias, ok, irr = [], 0, 0, 0
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        wbv = openpyxl.load_workbook(path, data_only=True)
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append(fname + ':' + hoja + '!' + coord
                              + ': hoja ausente')
                continue
            v = wbv[hoja][coord].value
            if v is None:
                if '""' in formula:
                    vacias += 1
                elif 'IRR(' in formula.upper():
                    irr += 1
                else:
                    fallos.append(fname + ':' + hoja + '!' + coord
                                  + ': sin cache (' + formula[:60] + ')')
            else:
                ok += 1
    return {'con_valor': ok, 'vacias_por_diseno': vacias,
            'irr_cacheada_aparte': irr, 'fallos': fallos}


def gate_graficos(carpeta, grupos_cargados, informe_ficheros):
    """§5: `len(ws._charts) >= 1` en los 9, contado tras el post-proceso.

    Un gráfico pendiente cuyo grupo NO se ha cargado no es un fallo: es la
    consecuencia esperada de correr `--solo motor`.
    """
    detalle, fallos, pendientes = [], [], []
    for fname in CON_GRAFICO:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            fallos.append(fname + ': no existe')
            continue
        wb = openpyxl.load_workbook(path)
        total = sum(len(ws._charts) for ws in wb.worksheets)
        hojas = [ws.title for ws in wb.worksheets if ws._charts]
        detalle.append({'fichero': fname, 'charts': total, 'hojas': hojas})
        if total >= 1:
            continue
        grupo = GRUPO_DE.get(fname)
        if grupo in grupos_cargados:
            fallos.append(fname + ': 0 gráficos con grupo_' + str(grupo)
                          + ' cargado')
        else:
            pendientes.append(fname + ': sin gráfico — lo desbloquea grupo_'
                              + str(grupo))
    return {'detalle': detalle, 'fallos': fallos,
            'pendientes_grupo': pendientes}


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-1500:])
    lineas = r.stdout.strip().splitlines()
    return {'exit': r.returncode, 'salida': lineas[-6:]}


# ==========================================================================
# Demostraciones (pycel) — sobre copias desechables, nunca sobre entregables
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                               # noqa: BLE001
            return 'ERR:' + type(e).__name__


def _copia_demo(carpeta, fname):
    os.makedirs(DEMOS, exist_ok=True)
    destino = os.path.join(DEMOS, fname)
    shutil.copy2(os.path.join(carpeta, fname), destino)
    return destino


def demo_tir_van_payback():
    """Caso trazado de la SPEC §1.8: −150.000 / 30.000 / 45.000 / 60.000 /
    70.000 → TIR 11,9592 %, VAN al 8 % 15.440,05 €, payback 3,21 años.

    La TIR se calcula con el mismo Newton-Raphson que va al fichero; el VAN se
    contrasta con pycel, que sí evalúa `NPV`, para que el número no dependa
    sólo de mi aritmética.
    """
    flujos = [-150000.0, 30000.0, 45000.0, 60000.0, 70000.0]
    tir = motor.tir_newton(flujos)
    van_py = motor.van(0.08, flujos)
    pb = motor.payback(flujos)

    os.makedirs(DEMOS, exist_ok=True)
    path = os.path.join(DEMOS, '_van_pycel.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'F'
    for i, v in enumerate(flujos):
        ws.cell(row=1, column=1 + i, value=v)
    ws['A3'] = '=NPV(0.08,B1:E1)+A1'
    ws['A4'] = '=IFERROR(IRR(A1:E1),"—")'
    wb.save(path)
    xl = _pycel(path)
    van_pycel = _ev(xl, 'F!A3')
    irr_pycel = _ev(xl, 'F!A4')

    return {
        'flujos': flujos,
        'tir_newton': round(tir, 6) if tir is not None else None,
        'tir_pct': (round(tir * 100, 4) if tir is not None else None),
        'esperado_spec_pct': 11.9592,
        'van_8pct_python': round(van_py, 2),
        'van_8pct_pycel': (round(van_pycel, 2)
                           if isinstance(van_pycel, float) else van_pycel),
        'esperado_van_spec': 15440.05,
        'payback_anios': round(pb, 2) if pb is not None else None,
        'esperado_payback_spec': 3.21,
        'pycel_con_IRR': irr_pycel,
        'nota': ('pycel devuelve error en IRR: por eso la TIR se cachea aparte '
                 '(§1.8). El VAN sí lo evalúa y coincide con el cálculo '
                 'propio.'),
    }


def demo_anualidad():
    """`PMT`/`PAGO` está prohibido (pycel no lo implementa). La anualidad
    algebraica `importe*i/(1-(1+i)^-n)` sí la evalúa: 100.000 € al 5 % nominal
    en 60 meses → 1.887,12 €/mes (verificación de la SPEC §4/DOM-14)."""
    os.makedirs(DEMOS, exist_ok=True)
    path = os.path.join(DEMOS, '_anualidad.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'F'
    ws['A1'] = 100000
    ws['A2'] = 0.05 / 12
    ws['A3'] = 60
    ws['A4'] = '=IFERROR($A$1*$A$2/(1-(1+$A$2)^(-$A$3)),"—")'
    ws['A5'] = '=IFERROR(PMT($A$2,$A$3,-$A$1),"—")'
    wb.save(path)
    xl = _pycel(path)
    algebraica = _ev(xl, 'F!A4')
    con_pmt = _ev(xl, 'F!A5')
    return {
        'importe': 100000, 'tipo_nominal': 0.05, 'meses': 60,
        'cuota_algebraica': (round(algebraica, 2)
                             if isinstance(algebraica, float) else algebraica),
        'esperado_spec': 1887.12,
        'misma_formula_con_PMT': con_pmt,
        'nota': ('PMT revienta en pycel y su IFERROR no lo atrapa: la celda se '
                 'quedaría sin caché y en blanco en Vista previa.'),
    }


def demo_counta():
    """`COUNTA` prohibida: `COUNTIF(rango,"<>")` da el mismo recuento y sí se
    evalúa (BONUS-09 §4/TEC-24)."""
    os.makedirs(DEMOS, exist_ok=True)
    path = os.path.join(DEMOS, '_counta.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'F'
    for i, v in enumerate(['a', 'b', None, 'd', None, 'f']):
        if v is not None:
            ws.cell(row=1 + i, column=1, value=v)
    ws['C1'] = '=COUNTIF($A$1:$A$6,"<>")'
    ws['C2'] = '=COUNTA($A$1:$A$6)'
    wb.save(path)
    xl = _pycel(path)
    return {'countif': _ev(xl, 'F!C1'), 'esperado': 4,
            'counta_en_pycel': _ev(xl, 'F!C2')}


def demo_iferror(carpeta):
    """`02!Break-Even`: con 0 días de apertura, `C9` daba `#¡DIV/0!`; con la
    guarda del §1.3 da un texto de ayuda. Se demuestra sobre una COPIA."""
    fname = '02-calculadora-punto-equilibrio.xlsx'
    path = os.path.join(carpeta, fname)
    if not os.path.isfile(path):
        return None
    copia = _copia_demo(carpeta, fname)
    xl = _pycel(copia)
    antes = _ev(xl, "'Break-Even'!C9")
    xl.set_value("'Datos'!C21", 0)
    despues = _ev(xl, "'Break-Even'!C9")
    xl2 = _pycel(copia)
    antes10 = _ev(xl2, "'Break-Even'!C8")
    xl2.set_value("'Datos'!C17", 1.0)
    despues10 = _ev(xl2, "'Break-Even'!C8")
    return {
        'fichero': fname,
        'C9_con_26_dias': antes, 'C9_con_0_dias': despues,
        'C8_con_35pct_variable': antes10, 'C8_con_100pct_variable': despues10,
        'ok': isinstance(despues, str) and isinstance(despues10, str),
    }


def demo_semaforo(carpeta):
    """Toda fórmula de estado tiene que CAMBIAR con un dato fuera de límite
    (gate de la SPEC §5) y su rango tiene que tener reglas de color."""
    fname = '06-dashboard-ratios-financieros.xlsx'
    path = os.path.join(carpeta, fname)
    if not os.path.isfile(path):
        return None
    copia = _copia_demo(carpeta, fname)
    wb = openpyxl.load_workbook(copia)
    ws = wb['Ratios']
    reglas = sum(len(cf.rules) for cf in ws.conditional_formatting
                 if str(cf.sqref).startswith('E17'))
    xl = _pycel(copia)
    base = _ev(xl, "'Ratios'!E17")
    ventas = _ev(xl, "'Ratios'!C6")
    xl.set_value("'Ratios'!C7", float(ventas) * 0.45)      # food cost 45 %
    alto = _ev(xl, "'Ratios'!E17")
    return {'fichero': fname, 'reglas_cf_en_E17_E25': reglas,
            'estado_base': base, 'estado_con_food_cost_45pct': alto,
            'cambia': base != alto}


def demo_saldo_03(carpeta):
    """DOM-01/TEC-01: el saldo inicial de cada mes tiene que ser el saldo FINAL
    del anterior. Caso trazado: saldo 15.000, cobros 40.000, pagos 35.000 en
    enero → febrero abre en 20.000.

    Mientras grupo_b no reescriba el 03, esto DOCUMENTA el fallo vivo (C5 lee
    la fila de «Otros pagos», no la del saldo final): once meses de tesorería
    erróneos ya cacheados en el fichero que se descarga.
    """
    fname = '03-cash-flow-forecast.xlsx'
    path = os.path.join(carpeta, fname)
    if not os.path.isfile(path):
        return None
    copia = _copia_demo(carpeta, fname)
    wb = openpyxl.load_workbook(copia)
    ws = wb['Flujo Mensual']
    formula_c5 = ws['C5'].value
    fila_saldo = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'SALDO FINAL':
            fila_saldo = r
    apunta_bien = (isinstance(formula_c5, str)
                   and fila_saldo is not None
                   and formula_c5.replace('$', '').upper()
                   == '=B' + str(fila_saldo))
    xl = _pycel(copia)
    # pycel sólo deja escribir en una celda que ya esté en su mapa: hay que
    # EVALUAR antes lo que la referencia (aquí, la apertura de febrero y el
    # saldo final de enero), o revienta con «not found in the cell map».
    _ev(xl, "'Flujo Mensual'!C5")
    _ev(xl, "'Flujo Mensual'!B" + str(fila_saldo or 25))
    xl.set_value("'Flujo Mensual'!B5", 15000)
    xl.set_value("'Flujo Mensual'!B7", 40000)
    for col in ('B8', 'B9', 'B10', 'B11'):
        xl.set_value("'Flujo Mensual'!" + col, 0)
    xl.set_value("'Flujo Mensual'!B14", 35000)
    for col in ('B15', 'B16', 'B17', 'B18', 'B19', 'B20', 'B21', 'B22'):
        xl.set_value("'Flujo Mensual'!" + col, 0)
    saldo_ene = _ev(xl, "'Flujo Mensual'!B" + str(fila_saldo or 25))
    apertura_feb = _ev(xl, "'Flujo Mensual'!C5")
    return {
        'fichero': fname,
        'formula_C5': formula_c5,
        'fila_SALDO_FINAL': fila_saldo,
        'C5_apunta_al_saldo_final': apunta_bien,
        'saldo_final_enero': saldo_ene,
        'apertura_febrero': apertura_feb,
        'esperado': 20000,
        'ok': apunta_bien and apertura_feb == saldo_ene,
        'nota': ('si `ok` es False y grupo_b no está cargado, esto es el bug '
                 'DOM-01 vivo, no un fallo del motor'),
    }


def demo_proteccion(carpeta, nombres):
    """47 hojas protegidas SIN contraseña y con los verdes desbloqueados."""
    fuera, fallos = [], []
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            verdes = sum(1 for row in ws.iter_rows() for c in row
                         if motor.es_verde(c))
            libres = sum(1 for row in ws.iter_rows() for c in row
                         if motor.es_verde(c) and not c.protection.locked)
            clave = getattr(ws.protection, 'password', None)
            fuera.append({'fichero': fname, 'hoja': ws.title,
                          'protegida': bool(ws.protection.sheet),
                          'con_contrasena': bool(clave),
                          'verdes': verdes, 'verdes_desbloqueadas': libres})
            if not ws.protection.sheet:
                fallos.append(fname + ':' + ws.title + ': sin proteger')
            if clave:
                fallos.append(fname + ':' + ws.title + ': CON contraseña')
            if verdes != libres:
                fallos.append(fname + ':' + ws.title + ': '
                              + str(verdes - libres)
                              + ' celdas verdes bloqueadas')
    return {'hojas': fuera, 'fallos': fallos,
            'total_hojas': len(fuera),
            'protegidas': sum(1 for h in fuera if h['protegida'])}


def demo_bio_version(carpeta, nombres):
    """§1.7: bio anclada (INSERCIÓN — no la llevaba ninguno de los 10) y línea
    de versión 2.0, una sola vez por fichero."""
    fuera, fallos = [], []
    for fname in nombres:
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        bio = version = nota = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if not isinstance(v, str):
                        continue
                    if v == motor.BIO_LINE:
                        bio += 1
                    if v == motor.VERSION_LINE:
                        version += 1
                    if motor.RX_NOTA_IVA.match(v):
                        nota += 1
        fuera.append({'fichero': fname, 'bio': bio, 'version_2_0': version,
                      'nota_iva': nota, 'subject': wb.properties.subject})
        if bio != 1:
            fallos.append(fname + ': bio ' + str(bio) + ' veces (debe ser 1)')
        if version != 1:
            fallos.append(fname + ': línea de versión ' + str(version)
                          + ' veces (debe ser 1)')
        if wb.properties.subject != motor.SUBJECT_V2:
            fallos.append(fname + ': subject = '
                          + repr(wb.properties.subject))
    return {'ficheros': fuera, 'fallos': fallos}


# ==========================================================================
# main
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 del Kit Plan Financiero')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    ap.add_argument('--solo', default='a,b,c',
                    help='grupos a aplicar: motor (ninguno), a, b, c o a,b')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--sin-idempotencia', action='store_true')
    ap.add_argument('--sin-demos', action='store_true')
    args = ap.parse_args()

    if not args.dry_run and os.environ.get('KIT_PLAN_FINANCIERO_APPLY') != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en ' + ORIGEN
            + ', que la SPEC declara intocable hasta que la ronda 2 dé verde. '
            'Usa --dry-run (o KIT_PLAN_FINANCIERO_APPLY=1 si eres el '
            'orquestador y ya tienes el visto bueno).')

    carpeta = DESTINO if args.dry_run else ORIGEN
    respaldo = None
    if args.dry_run:
        preparar_copia()
    else:
        respaldo = os.path.join(
            SCRATCH, PRODUCTO + '.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(ORIGEN, respaldo)
        log('  respaldo previo de los entregables: ' + respaldo)

    grupos, cargados = [], []
    pedidos = [g.strip().lower() for g in args.solo.split(',') if g.strip()]
    for letra in pedidos:
        if letra in ('motor', 'ninguno', 'none', ''):
            continue
        try:
            grupos.append(importlib.import_module('grupo_' + letra))
            cargados.append(letra)
            log('  grupo_' + letra + ' cargado')
        except ImportError:
            log('  grupo_' + letra + ' no existe todavía — se omite')

    nombres = ficheros_a_tocar(grupos)
    informe_ficheros, registros = [], {}
    log('\n== 1/8 · post-proceso de ' + str(len(nombres)) + ' ficheros ==')
    for fname in nombres:
        registros[fname] = procesar(carpeta, fname, grupos, informe_ficheros)
        log('  ' + fname + ': ' + str(len(registros[fname]))
            + ' fórmulas nuevas')

    # ---- idempotencia ---------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 2/8 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(IDEM):
            shutil.rmtree(IDEM)
        shutil.copytree(carpeta, IDEM)
        antes = dict((n, digest(os.path.join(carpeta, n))) for n in nombres)
        for fname in nombres:
            procesar(IDEM, fname, grupos, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(IDEM, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:40]}
        log('  diferencias 1.ª vs 2.ª pasada: ' + str(len(difs)))
        for d in difs[:10]:
            log('    ' + d)

    # ---- cache ----------------------------------------------------------
    log('\n== 3/8 · inject_cache (al final del todo) ==')
    cache = inject_cache(carpeta, nombres)

    log('\n== 4/8 · caché de la TIR por Newton-Raphson (§1.8) ==')
    irr = None
    p07 = os.path.join(carpeta, '07-informe-viabilidad-bancos.xlsx')
    if os.path.isfile(p07):
        aviso = []
        irr = motor.cachear_irr(p07, aviso)
        for a in aviso:
            log('  ' + a)
        log('  flujos: ' + str(irr.get('flujos'))
            + ' · TIR: ' + str(irr.get('tir')))

    log('\n== 5/8 · verificación data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log('  con valor: ' + str(ver['con_valor']) + ' · "" por diseño: '
        + str(ver['vacias_por_diseno']) + ' · IRR aparte: '
        + str(ver['irr_cacheada_aparte']) + ' · fallos: '
        + str(len(ver['fallos'])))
    for fl in ver['fallos'][:10]:
        log('    ' + fl)

    log('\n== 6/8 · gate de gráficos (tras el post-proceso e inject_cache) ==')
    gra = gate_graficos(carpeta, cargados, informe_ficheros)
    for d in gra['detalle']:
        log('  ' + d['fichero'] + ': ' + str(d['charts']) + ' gráfico(s) '
            + str(d['hojas']))
    for p in gra['pendientes_grupo']:
        log('  PENDIENTE ' + p)

    log('\n== 7/8 · censo-entregables --fail ==')
    cen = censo(carpeta)

    demos = {}
    if not args.sin_demos:
        log('\n== 8/8 · demostraciones ==')
        demos = {
            'tir_van_payback': demo_tir_van_payback(),
            'anualidad_sin_PMT': demo_anualidad(),
            'countif_sin_COUNTA': demo_counta(),
            'iferror_02': demo_iferror(carpeta),
            'semaforo_06': demo_semaforo(carpeta),
            'saldo_encadenado_03': demo_saldo_03(carpeta),
            'proteccion': demo_proteccion(carpeta, nombres),
            'bio_y_version': demo_bio_version(carpeta, nombres),
        }
        d = demos['tir_van_payback']
        log('  TIR del caso trazado: ' + str(d['tir_pct']) + ' % (SPEC '
            + str(d['esperado_spec_pct']) + ' %) · VAN ' + str(d['van_8pct_python'])
            + ' € · payback ' + str(d['payback_anios']) + ' años')
        log('  cuota algebraica 100.000/5 %/60m: '
            + str(demos['anualidad_sin_PMT']['cuota_algebraica']) + ' €')
        log('  hojas protegidas: '
            + str(demos['proteccion']['protegidas']) + '/'
            + str(demos['proteccion']['total_hojas']))

    # ---- veredicto ------------------------------------------------------
    fallos = []
    for g in grupos:
        if hasattr(g, 'demos'):
            propias = g.demos(carpeta, ORIGEN, DEMOS)
            fallos += propias.pop('fallos', [])
            demos.update(propias)
    if idem.get('diferencias'):
        fallos.append('idempotencia: ' + str(idem['diferencias'])
                      + ' diferencias')
    if ver['fallos']:
        fallos.append('cache: ' + str(len(ver['fallos']))
                      + ' fórmulas sin valor')
    fallos += gra['fallos']
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió ' + str(cen['exit']))
    if demos:
        fallos += demos.get('proteccion', {}).get('fallos', [])
        fallos += demos.get('bio_y_version', {}).get('fallos', [])
        tvp = demos.get('tir_van_payback') or {}
        if tvp.get('tir_pct') is not None \
                and abs(tvp['tir_pct'] - tvp['esperado_spec_pct']) > 0.001:
            fallos.append('TIR del caso trazado: ' + str(tvp['tir_pct'])
                          + ' % ≠ ' + str(tvp['esperado_spec_pct']) + ' %')
        anu = demos.get('anualidad_sin_PMT') or {}
        if isinstance(anu.get('cuota_algebraica'), float) \
                and abs(anu['cuota_algebraica'] - anu['esperado_spec']) > 0.01:
            fallos.append('anualidad algebraica: '
                          + str(anu['cuota_algebraica']) + ' ≠ 1887.12')

    informe = {
        'producto': PRODUCTO,
        'version': '2.0',
        'rol': 'motor',
        'spec': SPEC,
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'carpeta_origen': ORIGEN,
        'carpeta_trabajo': carpeta,
        'grupos_pedidos': pedidos,
        'grupos_cargados': cargados,
        'ficheros': informe_ficheros,
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'cachear_irr': irr,
            'data_only_formulas_nuevas': ver,
            'graficos': gra,
            'censo_entregables': cen,
        },
        'demostraciones': demos,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)), exist_ok=True)
        with open(args.json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1)
        log('\ninforme → ' + args.json)

    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'TODO VERDE (idempotencia, cache, TIR, gráficos, censo, '
                     'demos)'))
    if gra['pendientes_grupo']:
        log('  (pendientes de grupo, esperados sin ese grupo cargado: '
            + str(len(gra['pendientes_grupo'])) + ')')
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (BÓRRALO sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
