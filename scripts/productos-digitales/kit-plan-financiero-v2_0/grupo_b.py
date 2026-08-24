#!/usr/bin/env python3
"""
grupo_b.py — §3 de la SPEC: tesorería, equilibrio e inversión (02, 03, 04).

Qué arregla, id por id:

  · DOM-01 / TEC-01 / COM-01 / RD-02 / RT-05 / RC-07 (alta) — el saldo inicial
    de cada mes del 03 leía la fila «Otros pagos» en vez de «SALDO FINAL»:
    once meses de tesorería erróneos y once alertas rojas falsas, **ya
    cacheadas en el fichero que se descarga**. Con el mapa nuevo,
    `C5='=B29'` … `M5='=L29'`.
  · DOM-10 / DOM-23 / TEC-17 / COM-09 / COM-15 / RD-12 / RT-24 / RC-08 — hoja
    `Parámetros` nueva: % de cobro con tarjeta y días de abono, plazo de pago a
    proveedores, tipos de IVA y estacionalidad por mes. La «Liquidación de IVA»
    deja de ser una celda verde que se teclea a mano y calcula sola en el
    calendario de la AEAT (abril, julio, octubre; enero queda como input del 4T
    anterior). Así `Instrucciones!B8` pasa a ser cierta.
  · DOM-13 / TEC-08 / COM-20 / RD-18 / RC-22 — `02!Escenarios` repetía 19.000 €
    como constante y llevaba el `26` clavado dentro de la fórmula mientras
    `Break-Even` sí leía de `Datos`: dos verdades en el mismo libro.
  · TEC-23 / DOM-24 / RD-17 / RC-23 — la cuota del préstamo estaba DENTRO de
    TOTAL COSTES FIJOS, así que el «EBITDA» de Escenarios iba después del
    servicio de deuda y luego se comparaba con los benchmarks de EBITDA del 06
    y del 07. Sale a fila propia y `Break-Even` gana break-even **operativo** y
    break-even **de caja**. Y `Datos!B7` deja de contradecirse («bruto» con una
    nota que decía «incluye SS empresa»).
  · DOM-25 / TEC-16 / COM-27 / RD-16 / RT-23 / RC-14 — `C10` valía 51,10 y el
    formato `#,##0` enseñaba «51»: 51 × 22 € × 26 días = 29.172 €, **58,77 €
    por debajo** del break-even que la propia hoja acababa de calcular.
    `ROUNDUP`. Y la fila «Ticket medio necesario» que la ficha promete deja de
    faltar.
  · DOM-15 / COM-13 / RD-22 / RC-09 — el CAPEX no mencionaba el IVA ni una vez
    y la landing prometía «totales con y sin IVA»: quien firma los pedidos paga
    base + 21 %, y con 150.000 € de presupuesto el desembolso real es 181.500.
    Cabecera nueva con Base / IVA % / IVA (€) / Total con IVA / Real /
    Desviación / Coef. amortización / Dotación anual, y de ahí sale la fila
    «Amortización» del 07.
  · DOM-16 / RD-23 — hoja `Otros conceptos de apertura`: traspaso y fianza,
    constitución, **stock inicial** de despensa y bodega, nóminas y formación
    pre-apertura, marketing de lanzamiento, imprevistos y **fondo de maniobra**.
    El propio kit se lo pedía al usuario (`BONUS-09!C15`) y no le daba dónde
    ponerlo. `Licencias!A13` «Seguros (año 1)» se MUEVE aquí (§6: no se borra,
    que quien ya lo tenía presupuestado no lo pierda).
  · RD-24 — en la lista de compra de apertura no había ni una línea de vajilla,
    cristalería, cubertería, menaje ni uniformes/EPI: en 60 plazas con dos
    juegos de reposición son 4.000-8.000 € y es el olvido clásico que revienta
    el presupuesto la semana de la apertura.
  · RD-06 — `Escenarios` quedaba con CERO celdas editables porque sus inputs
    van pintados con los colores de escenario y `proteger()` desbloqueaba sólo
    lo verde. Se declaran editables por ROL (`motor.marcar_editable`).
  · RD-25 — el 03 se entregaba con once meses en «BAJO MÍNIMO» ya cacheados
    porque el único dato de ejemplo era el saldo inicial: un muro de rojo antes
    de teclear nada, que enseña a ignorar el semáforo. El estado de un mes sin
    movimientos es «—», no una alarma.
  · RD-30 / RT-22 — el SALDO INICIAL admite negativo: un negocio puede empezar
    el mes en descubierto, y ése es justo el caso que esta plantilla existe
    para detectar.
  · RD-13 — línea de comisiones de plataformas en el bloque de pagos del 03.
"""
import copy
import re

from openpyxl.styles import Alignment, Font

import motor

FICHEROS = [
    '02-calculadora-punto-equilibrio.xlsx',
    '03-cash-flow-forecast.xlsx',
    '04-presupuesto-inversion-capex.xlsx',
]

MESES = ('Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct',
         'Nov', 'Dic')
COLS = ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M')

# --------------------------------------------------------------------------
# 03 — mapa de filas nuevo de 'Flujo Mensual'
# --------------------------------------------------------------------------
F_SALDO_INI = 5
F_VENTAS = 12
F_TARJETA = 13
F_COBROS = 14
F_PROVEEDORES = 16
F_TOTAL_PAGOS = 27
F_FLUJO = 28
F_SALDO_FIN = 29
F_COMPRAS = 33
F_SS_DEVENGADA = 34
F_IVA_REP = 36
F_IVA_SOP = 37
F_SUGERIDAS = 39

COBROS = (
    (7, 'Ventas comedor (IVA incl.)'),
    (8, 'Ventas barra (IVA incl.)'),
    (9, 'Ventas delivery (IVA incl.)'),
    (10, 'Eventos / Catering (IVA incl.)'),
    (11, 'Otros cobros'),
)
# Filas 17-26: pagos con input directo (la 16, la 18 y la 25 son especiales).
PAGOS_INPUT = (
    (17, 'Nóminas (neto)'),
    (19, 'Alquiler'),
    (20, 'Suministros'),
    (21, 'Marketing'),
    (22, 'Gestoría / Admin'),
    (23, 'Cuota de préstamo'),
    (24, 'Retenciones IRPF (mod. 111)'),
    (26, 'Comisiones de plataformas de delivery y otros pagos'),
)

# Trimestres del modelo 303: la columna donde se PAGA y el rango de meses que
# se liquida. Abril paga el 1T, julio el 2T, octubre el 3T; el 4T se paga en
# enero, pero es el del año ANTERIOR, así que `B` queda como input verde.
TRIMESTRES = (('E', 'B', 'D'), ('H', 'E', 'G'), ('K', 'H', 'J'))



def _es_num(x):
    """pycel devuelve `int` cuando el resultado es entero: comprobar sólo
    `float` daba falsos fallos (121000 no es `float`)."""
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def _col(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


# ==========================================================================
# 03 · Previsión de tesorería
# ==========================================================================
def _parametros_03(wb, informe):
    if 'Parámetros' in wb.sheetnames:
        ws = wb['Parámetros']
    else:
        ws = wb.create_sheet('Parámetros', 1)
    motor.limpiar_rango(ws, 'A1:E40')
    ws['B1'] = 'Parámetros de Tesorería'
    ws['B1'].font = Font(bold=True, size=14)
    ws['B2'] = ('AI Chef Pro · aichef.pro — Kit Plan Financiero para '
                'Restaurantes')
    ws['B2'].font = Font(size=9)
    cab = ('B3', 'C3', 'D3')
    for coord, texto in zip(cab, ('Concepto', 'Valor', 'Notas')):
        ws[coord] = texto
        ws[coord].font = Font(bold=True, color='FFFFFF')
        ws[coord].fill = motor.PatternFill('solid', fgColor=motor.CAB)

    filas = (
        (4, '% de ventas cobradas con tarjeta', 0.65, motor.FMT_PCT0,
         'El efectivo entra el mismo día; la tarjeta, con desfase.'),
        (5, 'Días de abono de la tarjeta (D+n)', 2, motor.FMT_ENT,
         'Lo que tarda tu TPV en ingresarte. Típico D+1 o D+2.'),
        (6, 'Plazo de pago a proveedores (días)', 30, motor.FMT_ENT,
         'El que hayas negociado. 0 = pago al contado.'),
        (7, 'IVA repercutido en ventas (%)', 0.10, motor.FMT_PCT,
         'Restauración: 10 %.'),
        (8, 'IVA soportado en compras de alimentación (%)', 0.10,
         motor.FMT_PCT, 'Alimentación: 10 % (4 % en algunos básicos).'),
        (9, 'IVA soportado en el resto de gastos (%)', 0.21, motor.FMT_PCT,
         'Alquiler, suministros, marketing, gestoría: 21 %.'),
    )
    for fila, etiqueta, valor, fmt, nota in filas:
        ws['B' + str(fila)] = etiqueta
        motor.val(ws, 'C' + str(fila), valor, fmt, verde_=True)
        ws['D' + str(fila)] = nota
        ws['D' + str(fila)].font = Font(size=9, italic=True)

    ws['B11'] = 'ESTACIONALIDAD (índice por mes; 1,00 = mes tipo)'
    ws['B11'].font = Font(bold=True)
    for coord, texto in (('B12', 'Mes'), ('C12', 'Índice'),
                         ('D12', 'Ventas sugeridas del mes (IVA incl.)')):
        ws[coord] = texto
        ws[coord].font = Font(bold=True, color='FFFFFF')
        ws[coord].fill = motor.PatternFill('solid', fgColor=motor.CAB)
    for i, mes in enumerate(MESES):
        r = str(13 + i)
        ws['B' + r] = mes
        motor.val(ws, 'C' + r, 1.0, '0.00', verde_=True)
        motor.f(ws, 'D' + r, '=$C$26*$C' + r, motor.FMT_EUR)

    ws['B26'] = 'Ventas de un mes tipo (IVA incl.)'
    ws['B26'].font = Font(bold=True)
    motor.val(ws, 'C26', 0, motor.FMT_EUR, verde_=True)
    ws['D26'] = ('Escríbelo una vez y la estacionalidad te sugiere los doce '
                 'meses; luego cópialos a «Flujo Mensual».')
    ws['D26'].font = Font(size=9, italic=True)

    ws['B28'] = ('Calendario del modelo 303: se presenta del 1 al 20 de abril, '
                 'julio y octubre, y del 1 al 30 de enero. Por eso la '
                 'liquidación sólo calcula en esos meses.')
    ws['B28'].font = Font(size=9, italic=True)
    ws['A31'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A31'].font = Font(size=8)
    motor.anchos(ws, {'B': 46, 'C': 16, 'D': 58})
    informe.append("03: hoja 'Parámetros' con tarjeta, plazo de proveedores, "
                   'tipos de IVA y estacionalidad (DOM-10/DOM-23/COM-15)')


def _flujo_mensual(wb, informe):
    ws = wb['Flujo Mensual']
    # Fuentes de estilo ESTABLES: la fila 23 es «TOTAL PAGOS» en la v1.1 y
    # «Cuota de préstamo» (input) en la v2.0, así que copiar de ahí pintaba de
    # verde los totales en la 2.ª pasada. La 12 es un total en las dos.
    est_label = copy.copy(ws['A7']._style)
    est_input = copy.copy(ws['B7']._style)
    est_titulo = copy.copy(ws['A6']._style)
    est_total_lbl = copy.copy(ws['A12']._style)
    est_total_val = copy.copy(ws['B12']._style)

    motor.limpiar_rango(ws, 'A5:M45')

    # --- SALDO INICIAL: encadenado al SALDO FINAL del mes anterior ----------
    ws['A5'] = 'SALDO INICIAL'
    ws['A5']._style = copy.copy(est_total_lbl)
    motor.aplicar_estilo(ws, 'B5', est_input)
    motor.val(ws, 'B5', 15000, motor.FMT_EUR, verde_=True)
    motor.permitir_negativo(ws, 'B5')
    for i in range(1, 12):
        motor.aplicar_estilo(ws, COLS[i] + '5', est_total_val)
        motor.f(ws, COLS[i] + '5',
                '=' + COLS[i - 1] + str(F_SALDO_FIN), motor.FMT_EUR)

    def _bloque(fila, etiqueta, estilo=est_label):
        ws['A' + str(fila)] = etiqueta
        ws['A' + str(fila)]._style = copy.copy(estilo)

    def _inputs(fila):
        for L in COLS:
            coord = L + str(fila)
            motor.aplicar_estilo(ws, coord, est_input)
            motor.val(ws, coord, 0, motor.FMT_EUR, verde_=True)

    _bloque(6, 'COBROS (Entradas)', est_titulo)
    for fila, etiqueta in COBROS:
        _bloque(fila, etiqueta)
        _inputs(fila)

    _bloque(F_VENTAS, 'VENTAS DEL MES', est_total_lbl)
    _bloque(F_TARJETA, 'Pendiente de abono de tarjeta (a cobrar el mes que '
                       'viene)')
    _bloque(F_COBROS, 'TOTAL COBROS DE CAJA', est_total_lbl)
    for i, L in enumerate(COLS):
        motor.aplicar_estilo(ws, L + str(F_VENTAS), est_total_val)
        motor.f(ws, L + str(F_VENTAS), '=SUM(' + L + '7:' + L + '11)',
                motor.FMT_EUR)
        motor.aplicar_estilo(ws, L + str(F_COBROS), est_total_val)
        motor.f(ws, L + str(F_TARJETA),
                '=IFERROR(' + L + str(F_VENTAS)
                + '*Parámetros!$C$4*Parámetros!$C$5/30,0)', motor.FMT_EUR)
        if i == 0:
            motor.f(ws, L + str(F_COBROS),
                    '=' + L + str(F_VENTAS) + '-' + L + str(F_TARJETA),
                    motor.FMT_EUR)
        else:
            motor.f(ws, L + str(F_COBROS),
                    '=' + L + str(F_VENTAS) + '-' + L + str(F_TARJETA) + '+'
                    + COLS[i - 1] + str(F_TARJETA), motor.FMT_EUR)

    _bloque(15, 'PAGOS (Salidas)', est_titulo)
    _bloque(F_PROVEEDORES, 'Pago a proveedores (según el plazo negociado)')
    for i, L in enumerate(COLS):
        if i == 0:
            motor.f(ws, L + str(F_PROVEEDORES),
                    '=IFERROR(' + L + str(F_COMPRAS)
                    + '*(1-Parámetros!$C$6/30),0)', motor.FMT_EUR)
        else:
            motor.f(ws, L + str(F_PROVEEDORES),
                    '=IFERROR(' + L + str(F_COMPRAS)
                    + '*(1-Parámetros!$C$6/30)+'
                    + COLS[i - 1] + str(F_COMPRAS)
                    + '*Parámetros!$C$6/30,0)', motor.FMT_EUR)

    _bloque(18, 'Pago de Seguridad Social (la del mes anterior)')
    motor.aplicar_estilo(ws, 'B18', est_input)
    motor.val(ws, 'B18', 0, motor.FMT_EUR, verde_=True)
    for i in range(1, 12):
        motor.f(ws, COLS[i] + '18', '=' + COLS[i - 1] + str(F_SS_DEVENGADA),
                motor.FMT_EUR)

    for fila, etiqueta in PAGOS_INPUT:
        _bloque(fila, etiqueta)
        _inputs(fila)

    _bloque(25, 'Liquidación de IVA (mod. 303)')
    motor.aplicar_estilo(ws, 'B25', est_input)
    motor.val(ws, 'B25', 0, motor.FMT_EUR, verde_=True)
    calculadas = set()
    for destino, ini, fin in TRIMESTRES:
        motor.f(ws, destino + '25',
                '=MAX(0,SUM($' + ini + '$' + str(F_IVA_REP) + ':$' + fin + '$'
                + str(F_IVA_REP) + ')-SUM($' + ini + '$' + str(F_IVA_SOP)
                + ':$' + fin + '$' + str(F_IVA_SOP) + '))', motor.FMT_EUR)
        calculadas.add(destino)
    for L in COLS[1:]:
        if L not in calculadas:
            motor.f(ws, L + '25', '=""', motor.FMT_EUR)

    _bloque(F_TOTAL_PAGOS, 'TOTAL PAGOS', est_total_lbl)
    _bloque(F_FLUJO, 'FLUJO NETO DEL MES', est_total_lbl)
    _bloque(F_SALDO_FIN, 'SALDO FINAL', est_total_lbl)
    for L in COLS:
        for fila in (F_TOTAL_PAGOS, F_FLUJO, F_SALDO_FIN):
            motor.aplicar_estilo(ws, L + str(fila), est_total_val)
        motor.f(ws, L + str(F_TOTAL_PAGOS),
                '=SUM(' + L + '16:' + L + '26)', motor.FMT_EUR)
        motor.f(ws, L + str(F_FLUJO),
                '=' + L + str(F_COBROS) + '-' + L + str(F_TOTAL_PAGOS),
                motor.FMT_EUR)
        motor.f(ws, L + str(F_SALDO_FIN),
                '=' + L + '5+' + L + str(F_FLUJO), motor.FMT_EUR)

    # --- Bloque gris: datos base que NO son caja ----------------------------
    _bloque(31, 'DATOS BASE — no son caja, pero alimentan las filas de arriba',
            est_titulo)
    _bloque(F_COMPRAS, 'Compras de materia prima del mes (IVA incl.)')
    _bloque(F_SS_DEVENGADA, 'Seguridad Social devengada del mes')
    for fila in (F_COMPRAS, F_SS_DEVENGADA):
        _inputs(fila)
    _bloque(F_IVA_REP, 'IVA repercutido del mes')
    _bloque(F_IVA_SOP, 'IVA soportado del mes')
    _bloque(F_SUGERIDAS, 'Ventas sugeridas por estacionalidad (informativo)')
    for i, L in enumerate(COLS):
        motor.f(ws, L + str(F_IVA_REP),
                '=IFERROR(' + L + str(F_VENTAS) + '-' + L + str(F_VENTAS)
                + '/(1+Parámetros!$C$7),0)', motor.FMT_EUR)
        gastos = ('+'.join(L + str(n) for n in (19, 20, 21, 22)))
        motor.f(ws, L + str(F_IVA_SOP),
                '=IFERROR(' + L + str(F_COMPRAS) + '-' + L + str(F_COMPRAS)
                + '/(1+Parámetros!$C$8)+(' + gastos + ')-(' + gastos
                + ')/(1+Parámetros!$C$9),0)', motor.FMT_EUR)
        motor.f(ws, L + str(F_SUGERIDAS),
                '=Parámetros!$D$' + str(13 + i), motor.FMT_EUR)
    # OJO: el gris NO puede pisar los inputs de las filas 33 y 34 — el relleno
    # es lo que `proteger()` mira para decidir qué queda editable, así que
    # pintarlas de gris las dejaría bloqueadas.
    motor.gris(ws, 'A31:M31')
    motor.gris(ws, 'A33:A34')
    motor.gris(ws, 'A' + str(F_IVA_REP) + ':M' + str(F_IVA_SOP))
    motor.gris(ws, 'A' + str(F_SUGERIDAS) + ':M' + str(F_SUGERIDAS))

    ws['A41'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A41'].font = Font(size=8)
    motor.anchos(ws, {'A': 46})
    informe.append('03: mapa nuevo de «Flujo Mensual» — saldo encadenado a la '
                   'fila ' + str(F_SALDO_FIN) + ', desfase de tarjeta, plazo '
                   'de proveedores, SS del mes siguiente e IVA trimestral '
                   'calculado (DOM-01/TEC-01/DOM-10/DOM-23)')


def _alertas_03(wb, informe):
    ws = wb['Alertas']
    for i, L in enumerate(COLS):
        fila = str(6 + i)
        motor.f(ws, 'C' + fila, "='Flujo Mensual'!" + L + str(F_SALDO_FIN),
                motor.FMT_EUR)
        # RD-25: un mes SIN movimientos no opina. El fichero salía con once
        # «BAJO MÍNIMO» cacheados sólo porque todo estaba a cero.
        motor.f(ws, 'D' + fila,
                "=IF('Flujo Mensual'!" + L + str(F_VENTAS)
                + "+'Flujo Mensual'!" + L + str(F_TOTAL_PAGOS)
                + '=0,"—",IF(C' + fila + '<$C$3,"⚠️ BAJO MÍNIMO","OK"))')
    ws['B19'] = ('El estado «—» significa que ese mes todavía no tiene ni '
                 'cobros ni pagos: el semáforo no opina sobre un mes vacío.')
    ws['B19'].font = Font(size=9, italic=True)
    informe.append('03: Alertas leen el SALDO FINAL de la fila '
                   + str(F_SALDO_FIN) + ' y callan en los meses sin datos '
                   '(RD-25)')


# ==========================================================================
# 02 · Punto de equilibrio
# ==========================================================================
DATOS_02 = (
    (6, 'Alquiler mensual', 3000, motor.FMT_EUR, '€/mes'),
    (7, 'Coste total de personal fijo (salario bruto + SS empresa)', 8800,
     motor.FMT_EUR, '€/mes — ≈ bruto × 1,32; es coste de EMPRESA, no bruto'),
    (8, 'Seguros', 400, motor.FMT_EUR, '€/mes'),
    (9, 'Suministros (luz, agua, gas)', 1500, motor.FMT_EUR, '€/mes'),
    (10, 'Marketing fijo', 500, motor.FMT_EUR, '€/mes'),
    (11, 'Gestoría / Administración', 300, motor.FMT_EUR, '€/mes'),
    (13, 'Otros gastos fijos', 500, motor.FMT_EUR, '€/mes'),
)


def _datos_02(wb, informe):
    ws = wb['Datos']
    est_label = copy.copy(ws['B6']._style)
    est_input = copy.copy(ws['C6']._style)
    est_nota = copy.copy(ws['D6']._style)
    est_total_lbl = copy.copy(ws['B14']._style)
    est_total_val = copy.copy(ws['C14']._style)

    motor.limpiar_rango(ws, 'B6:D24')

    for fila, etiqueta, valor, fmt, nota in DATOS_02:
        r = str(fila)
        ws['B' + r] = etiqueta
        ws['B' + r]._style = copy.copy(est_label)
        motor.aplicar_estilo(ws, 'C' + r, est_input)
        motor.val(ws, 'C' + r, valor, fmt, verde_=True)
        motor.aplicar_estilo(ws, 'D' + r, est_nota)
        ws['D' + r] = nota

    ws['B12'] = ('La cuota del préstamo NO es un coste fijo operativo: va en '
                 'la fila 15, fuera del EBITDA.')
    ws['B12'].font = Font(size=9, italic=True)

    ws['B14'] = 'TOTAL COSTES FIJOS OPERATIVOS'
    ws['B14']._style = copy.copy(est_total_lbl)
    motor.aplicar_estilo(ws, 'C14', est_total_val)
    motor.f(ws, 'C14', '=SUM(C6:C11)+C13', motor.FMT_EUR)

    ws['B15'] = 'Cuota de préstamo (servicio de deuda)'
    ws['B15']._style = copy.copy(est_label)
    motor.aplicar_estilo(ws, 'C15', est_input)
    motor.val(ws, 'C15', 800, motor.FMT_EUR, verde_=True)
    motor.aplicar_estilo(ws, 'D15', est_nota)
    ws['D15'] = 'Cópiala del cuadro de amortización: 07!Financiación'

    ws['B16'] = 'COSTES VARIABLES'
    ws['B16'].font = Font(bold=True)
    ws['B17'] = '% Coste variable sobre ventas'
    ws['B17']._style = copy.copy(est_label)
    motor.aplicar_estilo(ws, 'C17', est_input, motor.FMT_PCT)
    motor.val(ws, 'C17', 0.35, motor.FMT_PCT, verde_=True)
    motor.aplicar_estilo(ws, 'D17', est_nota)
    ws['D17'] = ('Food cost + comisiones de delivery + variables (30-40 % '
                 'típico). NO incluye el personal fijo.')

    ws['B19'] = 'TICKET MEDIO'
    ws['B19'].font = Font(bold=True)
    ws['B20'] = 'Ticket medio por comensal (€, sin IVA)'
    ws['B20']._style = copy.copy(est_label)
    motor.val(ws, 'C20', motor.EJEMPLO_CANONICO['ticket'], motor.FMT_EUR,
              verde_=True)
    ws['B21'] = 'Días de apertura al mes'
    ws['B21']._style = copy.copy(est_label)
    motor.val(ws, 'C21', motor.EJEMPLO_CANONICO['dias_apertura'],
              motor.FMT_ENT, verde_=True)
    ws['A24'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A24'].font = Font(size=8)
    informe.append('02: la cuota sale del TOTAL COSTES FIJOS y el personal '
                   'tiene una etiqueta única (TEC-23/DOM-24); ejemplo '
                   'recalibrado a la base común del kit (§7.5)')


def _break_even_02(wb, informe):
    ws = wb['Break-Even']
    est_label = copy.copy(ws['B5']._style)
    est_val = copy.copy(ws['C5']._style)
    est_total_lbl = copy.copy(ws['B8']._style)
    est_total_val = copy.copy(ws['C8']._style)

    motor.limpiar_rango(ws, 'A5:F32')

    def _fila(r, etiqueta, formula, fmt, fuerte=False):
        ws['B' + str(r)]._style = copy.copy(
            est_total_lbl if fuerte else est_label)
        ws['B' + str(r)] = etiqueta
        motor.aplicar_estilo(ws, 'C' + str(r),
                             est_total_val if fuerte else est_val, fmt)
        motor.f(ws, 'C' + str(r), formula, fmt)

    _fila(5, 'Total costes fijos operativos / mes', '=Datos!$C$14',
          motor.FMT_EUR)
    _fila(6, '% Coste variable', '=Datos!$C$17', motor.FMT_PCT)
    _fila(7, 'Margen de contribución', '=1-Datos!$C$17', motor.FMT_PCT)
    _fila(8, 'FACTURACIÓN BREAK-EVEN / mes',
          motor.iferror('Datos!$C$14/(1-Datos!$C$17)',
                        'Revisa el % de coste variable (no puede ser 100 %)'),
          motor.FMT_EUR, fuerte=True)
    _fila(9, 'Facturación break-even / día',
          motor.iferror('$C$8/Datos!$C$21', 'Indica los días de apertura'),
          motor.FMT_EUR)
    # DOM-25: ROUNDUP. Con 51,10 y formato #,##0 el usuario leía «51» y con 51
    # cubiertos NO llega al punto de equilibrio.
    _fila(10, 'Cubiertos / día necesarios',
          motor.iferror('ROUNDUP($C$9/Datos!$C$20,0)',
                        'Indica el ticket medio'), motor.FMT_ENT, fuerte=True)
    _fila(11, 'Cubiertos / mes necesarios',
          motor.iferror('ROUNDUP($C$8/Datos!$C$20,0)', '—'), motor.FMT_ENT)

    ws['B12'] = 'Cubiertos / día previstos'
    ws['B12']._style = copy.copy(est_label)
    motor.val(ws, 'C12', motor.EJEMPLO_CANONICO['cubiertos_dia'],
              motor.FMT_ENT, verde_=True)
    _fila(13, 'Ticket medio necesario (€)',
          motor.iferror('$C$9/$C$12', 'Indica los cubiertos/día previstos'),
          motor.FMT_EUR)

    ws['B15'] = 'BREAK-EVEN DE CAJA (incluye la cuota del préstamo)'
    ws['B15'].font = Font(bold=True)
    _fila(16, 'Facturación break-even de caja / mes',
          motor.iferror('(Datos!$C$14+Datos!$C$15)/(1-Datos!$C$17)',
                        'Revisa el % de coste variable (no puede ser 100 %)'),
          motor.FMT_EUR, fuerte=True)
    _fila(17, 'Cubiertos / día para el break-even de caja',
          motor.iferror('ROUNDUP($C$16/Datos!$C$21/Datos!$C$20,0)',
                        'Indica el ticket medio'), motor.FMT_ENT)

    # --- bloque auxiliar del gráfico (§1.1) --------------------------------
    ws['A19'] = 'BLOQUE AUXILIAR DEL GRÁFICO — no lo edites'
    ws['A19'].font = Font(bold=True, size=9)
    for coord, texto in (('A20', 'Cubiertos / día'), ('B20', 'Ingresos'),
                         ('C20', 'Costes totales'),
                         ('D20', 'Facturación break-even')):
        ws[coord] = texto
        ws[coord].font = Font(bold=True, color='FFFFFF')
        ws[coord].fill = motor.PatternFill('solid', fgColor=motor.CAB)
    for i in range(8):
        r = str(21 + i)
        # El multiplicador va como CONSTANTE en una columna auxiliar: §1.3
        # prohíbe literales dentro de la fórmula.
        motor.val(ws, 'F' + r, round(0.25 * (i + 1), 2), '0.00')
        motor.f(ws, 'A' + r, '=IFERROR(ROUND($C$10*$F' + r + ',0),0)',
                motor.FMT_ENT)
        motor.f(ws, 'B' + r,
                '=IFERROR($A' + r + '*Datos!$C$20*Datos!$C$21,0)',
                motor.FMT_EUR)
        motor.f(ws, 'C' + r,
                '=IFERROR(Datos!$C$14+$B' + r + '*Datos!$C$17,0)',
                motor.FMT_EUR)
        motor.f(ws, 'D' + r, '=$C$8', motor.FMT_EUR)
    motor.gris(ws, 'A20:F28')
    ws['A30'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A30'].font = Font(size=8)
    motor.anchos(ws, {'B': 44, 'C': 18})
    informe.append('02: ROUNDUP en el umbral, ticket medio necesario, '
                   'break-even de caja y bloque auxiliar del gráfico '
                   '(DOM-25/TEC-16/COM-27/§1.1)')


def _escenarios_02(wb, informe):
    ws = wb['Escenarios']
    # El pie © venía combinado en A16:E16 y ahí van ahora el resultado después
    # de deuda y el margen: `limpiar_rango` deshace la combinación (una
    # `MergedCell` tiene el valor de sólo lectura).
    motor.limpiar_rango(ws, 'A14:E22')
    ws['B6'] = ('% Coste variable (food + comisiones; NO incluye personal '
                'fijo)')
    for L in ('C', 'D', 'E'):
        # DOM-13: los costes fijos dejan de ser tres constantes de 19.000 €.
        motor.f(ws, L + '7', '=Datos!$C$14', motor.FMT_EUR)
        # TEC-08: los días de apertura salen de Datos, no del `26` literal.
        motor.f(ws, L + '10', '=' + L + '4*' + L + '5*Datos!$C$21',
                motor.FMT_EUR)
    ws['B7'] = 'Costes fijos operativos / mes (€) — de Datos!C14'
    # RD-06: los inputs van pintados con los colores de escenario, no de
    # verde: se declaran editables por ROL o la hoja queda sin una sola celda
    # que se pueda tocar.
    motor.marcar_editable(ws, 'C4:E6')

    ws['B14'] = 'Cuota de préstamo (servicio de deuda)'
    ws['B15'] = 'RESULTADO DESPUÉS DE DEUDA'
    ws['B15'].font = Font(bold=True)
    ws['B16'] = 'Margen EBITDA %'
    for L in ('C', 'D', 'E'):
        motor.f(ws, L + '14', '=Datos!$C$15', motor.FMT_EUR)
        motor.f(ws, L + '15', '=' + L + '13-' + L + '14', motor.FMT_EUR)
        motor.f(ws, L + '16', '=IF(' + L + '10=0,0,' + L + '13/' + L + '10)',
                motor.FMT_PCT)
    ws['B18'] = ('El EBITDA de la fila 13 va ANTES del servicio de deuda: es '
                 'el que se compara con los benchmarks del 06 y del 07.')
    ws['B18'].font = Font(size=9, italic=True)
    ws['A20'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A20'].font = Font(size=8)
    motor.anchos(ws, {'B': 52})
    informe.append('02: Escenarios lee de Datos y sus inputs vuelven a ser '
                   'editables (DOM-13/TEC-08/COM-20/RD-06)')


# ==========================================================================
# 04 · CAPEX
# ==========================================================================
CAB_04 = ('Partida', 'Base (€)', 'IVA %', 'IVA (€)', 'Total con IVA (€)',
          'Real (€)', 'Desviación %', 'Coef. amortización',
          'Dotación anual (€)', 'Notas')

PARTIDAS_04 = {
    'Obra': (0.03, 0.21, (
        'Demolición y vaciado', 'Albañilería', 'Fontanería', 'Electricidad',
        'Climatización / Extracción', 'Pintura y acabados', 'Suelo',
        'Fachada / Rótulo', 'Baños clientes', 'Vestuarios personal')),
    'Equipamiento Cocina': (0.12, 0.21, (
        'Cocina industrial (fogones)', 'Horno combinado', 'Plancha / Grill',
        'Freidora', 'Cámara frigorífica', 'Congelador',
        'Mesa de trabajo inox', 'Lavavajillas industrial',
        'Campana extractora', 'Pequeño material cocina',
        'Menaje y utillaje de cocina (gastronorms, cuchillería)',
        'Estantería de almacén y cámara de bebidas')),
    'Mobiliario Sala': (0.10, 0.21, (
        'Mesas', 'Sillas', 'Barra', 'Taburetes barra',
        'Estanterías / decoración', 'Iluminación', 'Terraza (mesas + sillas)',
        'Textil (manteles, servilletas)',
        'Vajilla, cristalería y cubertería (con reposición)',
        'Uniformes y EPI')),
    'Tecnología': (0.25, 0.21, (
        'TPV (hardware + software)', 'Impresoras comandas', 'Datáfono',
        'Sistema de reservas', 'Wifi / Red', 'Pantallas / Digital signage',
        'Web / App delivery', 'Sistema de seguridad (cámaras)')),
    'Licencias': (0.0, 0.0, (
        'Licencia de apertura', 'Licencia de obras',
        'Proyecto técnico (arquitecto)', 'Certificado energético',
        'Plan de autoprotección', 'Alta Hacienda / SS', 'Registro sanitario',
        'SGAE / Derechos música')),
}

HOJA_APERTURA = 'Otros conceptos de apertura'
APERTURA = (
    ('Traspaso del local', 0.21, ''),
    ('Fianza y primeras rentas', 0.0,
     'La fianza se recupera al salir; las rentas anticipadas, no.'),
    ('Constitución, notaría y registro', 0.21, ''),
    ('STOCK INICIAL de despensa y bodega', 0.10,
     'Semanas de consumo × coste diario de materia prima. En un casual de 60 '
     'plazas, 2-4 semanas suelen ser 8.000-15.000 €.'),
    ('Nóminas y formación pre-apertura', 0.0,
     'Las dos o tres semanas de plantilla antes de facturar el primer euro.'),
    ('Marketing de lanzamiento', 0.21, ''),
    ('Seguros (año 1)', 0.0,
     'Gasto anticipado NO amortizable. Venía de la pestaña Licencias: no es '
     'CAPEX y duplicaba 02!Datos!C8.'),
    ('Imprevistos (5-10 % de la inversión)', 0.21,
     'Sobre «INVERSIÓN TOTAL (CAPEX)» de la pestaña Resumen.'),
    ('FONDO DE MANIOBRA (3-6 meses de costes fijos)', 0.0,
     'Es lo que pide BONUS-09!C15 y hasta hoy no tenía dónde ponerse. '
     'Referencia: 02!Datos!C14 × 3 a 6.'),
)


def _cabecera(ws, fila, textos):
    for i, texto in enumerate(textos):
        cel = ws.cell(row=fila, column=1 + i)
        cel.value = texto
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.fill = motor.PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center', wrap_text=True)


def _hoja_partidas(ws, coef, iva, partidas, informe):
    motor.limpiar_rango(ws, 'A4:J40')
    _cabecera(ws, 4, CAB_04)
    n = len(partidas)
    for i, nombre in enumerate(partidas):
        r = str(5 + i)
        ws['A' + r] = nombre
        motor.val(ws, 'B' + r, 0, motor.FMT_EUR, verde_=True)
        motor.val(ws, 'C' + r, iva, motor.FMT_PCT0, verde_=True)
        motor.f(ws, 'D' + r, '=$B' + r + '*$C' + r, motor.FMT_EUR)
        motor.f(ws, 'E' + r, '=$B' + r + '+$D' + r, motor.FMT_EUR)
        motor.val(ws, 'F' + r, 0, motor.FMT_EUR, verde_=True)
        motor.f(ws, 'G' + r,
                motor.iferror('($F' + r + '-$B' + r + ')/$B' + r),
                motor.FMT_PCT)
        motor.val(ws, 'H' + r, coef, motor.FMT_PCT, verde_=True)
        motor.f(ws, 'I' + r, '=$B' + r + '*$H' + r, motor.FMT_EUR)
        motor.val(ws, 'J' + r, None, None, verde_=True)
    t = 5 + n
    ws['A' + str(t)] = 'TOTAL ' + ws.title.upper()
    ws['A' + str(t)].font = Font(bold=True)
    for L in ('B', 'D', 'E', 'F', 'I'):
        motor.f(ws, L + str(t), '=SUM(' + L + '5:' + L + str(t - 1) + ')',
                motor.FMT_EUR)
        ws[L + str(t)].font = Font(bold=True)
    motor.f(ws, 'G' + str(t),
            motor.iferror('($F' + str(t) + '-$B' + str(t) + ')/$B' + str(t)),
            motor.FMT_PCT)
    ws['A' + str(t + 2)] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A' + str(t + 2)].font = Font(size=8)
    motor.anchos(ws, {'A': 44, 'B': 15, 'C': 9, 'D': 14, 'E': 17, 'F': 15,
                      'G': 13, 'H': 15, 'I': 17, 'J': 34})
    informe.append('04!' + ws.title + ': base, IVA, total con IVA y '
                   'coeficiente de amortización en ' + str(n) + ' partidas')
    return t


def _hoja_apertura(wb, informe):
    if HOJA_APERTURA in wb.sheetnames:
        ws = wb[HOJA_APERTURA]
    else:
        ws = wb.create_sheet(HOJA_APERTURA, len(wb.sheetnames) - 1)
    motor.limpiar_rango(ws, 'A1:H30')
    ws['A1'] = 'Otros conceptos de apertura — lo que no es CAPEX y hay que pagar igual'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = ('AI Chef Pro · aichef.pro — Kit Plan Financiero para '
                'Restaurantes')
    ws['A2'].font = Font(size=9)
    _cabecera(ws, 4, ('Concepto', 'Base (€)', 'IVA %', 'IVA (€)',
                      'Total con IVA (€)', 'Real (€)', 'Desviación %',
                      'Notas'))
    for i, (nombre, iva, nota) in enumerate(APERTURA):
        r = str(5 + i)
        ws['A' + r] = nombre
        motor.val(ws, 'B' + r, 0, motor.FMT_EUR, verde_=True)
        motor.val(ws, 'C' + r, iva, motor.FMT_PCT0, verde_=True)
        motor.f(ws, 'D' + r, '=$B' + r + '*$C' + r, motor.FMT_EUR)
        motor.f(ws, 'E' + r, '=$B' + r + '+$D' + r, motor.FMT_EUR)
        motor.val(ws, 'F' + r, 0, motor.FMT_EUR, verde_=True)
        motor.f(ws, 'G' + r,
                motor.iferror('($F' + r + '-$B' + r + ')/$B' + r),
                motor.FMT_PCT)
        ws['H' + r] = nota
        ws['H' + r].font = Font(size=9, italic=True)
        ws['H' + r].alignment = Alignment(wrap_text=True, vertical='top')
    t = 5 + len(APERTURA)
    ws['A' + str(t)] = 'TOTAL OTROS CONCEPTOS DE APERTURA'
    ws['A' + str(t)].font = Font(bold=True)
    for L in ('B', 'D', 'E', 'F'):
        motor.f(ws, L + str(t), '=SUM(' + L + '5:' + L + str(t - 1) + ')',
                motor.FMT_EUR)
        ws[L + str(t)].font = Font(bold=True)
    motor.f(ws, 'G' + str(t),
            motor.iferror('($F' + str(t) + '-$B' + str(t) + ')/$B' + str(t)),
            motor.FMT_PCT)
    ws['A' + str(t + 2)] = (
        'Ninguno de estos conceptos se amortiza: son gasto del ejercicio o '
        'circulante. Por eso no llevan coeficiente.')
    ws['A' + str(t + 2)].font = Font(size=9, italic=True)
    ws['A' + str(t + 4)] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A' + str(t + 4)].font = Font(size=8)
    motor.anchos(ws, {'A': 46, 'B': 15, 'C': 9, 'D': 14, 'E': 17, 'F': 15,
                      'G': 13, 'H': 56})
    informe.append("04: hoja '" + HOJA_APERTURA + "' con stock inicial, "
                   'fianza, imprevistos y fondo de maniobra (DOM-16/RD-23)')
    return t


def _resumen_04(wb, totales, total_apertura, informe):
    ws = wb['Resumen']
    motor.limpiar_rango(ws, 'A4:H32')
    _cabecera(ws, 4, ('Categoría', 'Base (€)', 'IVA (€)', 'Total con IVA (€)',
                      'Real (€)', 'Desviación %', 'Dotación anual (€)'))
    orden = ('Obra', 'Equipamiento Cocina', 'Mobiliario Sala', 'Tecnología',
             'Licencias')
    for i, hoja in enumerate(orden):
        r = str(5 + i)
        t = str(totales[hoja])
        ref = "='" + hoja + "'!"
        ws['A' + r] = hoja
        motor.f(ws, 'B' + r, ref + 'B' + t, motor.FMT_EUR)
        motor.f(ws, 'C' + r, ref + 'D' + t, motor.FMT_EUR)
        motor.f(ws, 'D' + r, ref + 'E' + t, motor.FMT_EUR)
        motor.f(ws, 'E' + r, ref + 'F' + t, motor.FMT_EUR)
        motor.f(ws, 'F' + r, motor.iferror('($E' + r + '-$B' + r + ')/$B' + r),
                motor.FMT_PCT)
        motor.f(ws, 'G' + r, ref + 'I' + t, motor.FMT_EUR)
    ws['A10'] = 'INVERSIÓN TOTAL (CAPEX)'
    ws['A10'].font = Font(bold=True)
    for L in ('B', 'C', 'D', 'E', 'G'):
        motor.f(ws, L + '10', '=SUM(' + L + '5:' + L + '9)', motor.FMT_EUR)
        ws[L + '10'].font = Font(bold=True)
    motor.f(ws, 'F10', motor.iferror('($E10-$B10)/$B10'), motor.FMT_PCT)

    ta = str(total_apertura)
    ws['A11'] = 'Otros conceptos de apertura'
    ref = "='" + HOJA_APERTURA + "'!"
    motor.f(ws, 'B11', ref + 'B' + ta, motor.FMT_EUR)
    motor.f(ws, 'C11', ref + 'D' + ta, motor.FMT_EUR)
    motor.f(ws, 'D11', ref + 'E' + ta, motor.FMT_EUR)
    motor.f(ws, 'E11', ref + 'F' + ta, motor.FMT_EUR)
    motor.f(ws, 'F11', motor.iferror('($E11-$B11)/$B11'), motor.FMT_PCT)

    ws['A12'] = 'DESEMBOLSO TOTAL DE APERTURA (base + IVA)'
    ws['A12'].font = Font(bold=True)
    for L in ('B', 'C', 'D', 'E'):
        motor.f(ws, L + '12', '=' + L + '10+' + L + '11', motor.FMT_EUR)
        ws[L + '12'].font = Font(bold=True)
    motor.f(ws, 'F12', motor.iferror('($E12-$B12)/$B12'), motor.FMT_PCT)

    ws['A14'] = 'De dónde sale cada dato (el encadenado realista entre libros):'
    ws['A14'].font = Font(bold=True, size=9)
    ws['A15'] = ('▸ La «Dotación anual (€)» de la fila 10 es la que se copia a '
                 '07!Proyecciones, fila «Amortización».')
    ws['A16'] = ('▸ El «IVA (€)» de la fila 12 es IVA soportado: se recupera '
                 'vía modelo 303, pero hay que ADELANTARLO. El préstamo tiene '
                 'que cubrirlo.')
    for r in ('A15', 'A16'):
        ws[r].font = Font(size=9)
    ws['A30'] = '© 2026 AI Chef Pro · aichef.pro'
    ws['A30'].font = Font(size=8)
    motor.anchos(ws, {'A': 40, 'B': 16, 'C': 14, 'D': 18, 'E': 16, 'F': 14,
                      'G': 18})
    informe.append('04!Resumen: base, IVA soportado, desembolso total y '
                   'dotación anual a la amortización (DOM-15/COM-13)')


# ==========================================================================
# API del grupo
# ==========================================================================
def post(wb, fname, cambios, registro):
    if fname.startswith('02'):
        _datos_02(wb, cambios)
        _break_even_02(wb, cambios)
        _escenarios_02(wb, cambios)
        ws = wb['Instrucciones']
        motor.linea_instrucciones(
            ws, "▸ En 'Datos' están los costes fijos OPERATIVOS. La cuota del "
                'préstamo va aparte (fila 15): el EBITDA se mide antes del '
                'servicio de deuda.',
            rx=re.compile(r"^▸ En la pestaña 'Datos'"))
        motor.linea_instrucciones(
            ws, "▸ 'Break-Even' da dos umbrales: el operativo (EBITDA = 0) y "
                'el de caja (con la cuota dentro), además del ticket medio '
                'necesario para tus cubiertos previstos.',
            rx=re.compile(r"^▸ La pestaña 'Break-Even'"))
        motor.linea_instrucciones(
            ws, '▸ Costes fijos: alquiler, personal fijo, seguros, gestoría — '
                'no varían con la actividad. La cuota del préstamo NO entra: '
                'es financiación, no explotación.',
            rx=re.compile(r'^▸ Costes Fijos:'))
    elif fname.startswith('03'):
        _parametros_03(wb, cambios)
        _flujo_mensual(wb, cambios)
        _alertas_03(wb, cambios)
        ws = wb['Instrucciones']
        motor.linea_instrucciones(
            ws, "▸ El IVA trimestral se calcula solo con los tipos de "
                "'Parámetros' y el calendario del modelo 303 (abril, julio y "
                'octubre); enero queda como input, porque liquida el 4T del '
                'año anterior.',
            rx=re.compile(r'^▸ El IVA trimestral'))
        motor.linea_instrucciones(
            ws, "▸ En 'Parámetros' pones el % de cobro con tarjeta y su "
                'desfase, el plazo de pago a proveedores y la estacionalidad: '
                'la tesorería deja de ser una lista de deseos.')
        motor.linea_instrucciones(
            ws, '▸ El bloque gris del final («datos base») no es caja: son las '
                'compras y la Seguridad Social devengada del mes, de donde '
                'salen el pago a proveedores, la SS del mes siguiente y el '
                'IVA.')
    elif fname.startswith('04'):
        totales = {}
        for hoja, (coef, iva, partidas) in PARTIDAS_04.items():
            if hoja in wb.sheetnames:
                totales[hoja] = _hoja_partidas(wb[hoja], coef, iva, partidas,
                                               cambios)
        total_apertura = _hoja_apertura(wb, cambios)
        _resumen_04(wb, totales, total_apertura, cambios)
        ws = wb['Instrucciones']
        motor.linea_instrucciones(
            ws, '▸ Cada partida lleva Base, IVA % e IVA (€): el desembolso '
                'real es la columna «Total con IVA», que con un 21 % es un '
                'quinto más de lo que enseña el presupuesto.',
            rx=re.compile(r'^▸ Introduce el presupuesto'))
        motor.linea_instrucciones(
            ws, '▸ El coeficiente de amortización por pestaña (obra 3 %, '
                'cocina 12 %, mobiliario 10 %, tecnología 25 %) da la '
                'dotación anual que pide el 07.')
        motor.linea_instrucciones(
            ws, "▸ Lo que no es CAPEX —traspaso, fianza, stock inicial, "
                'nóminas pre-apertura, imprevistos y fondo de maniobra— va en '
                "la pestaña 'Otros conceptos de apertura'.")
    return cambios


def demos(carpeta, origen, destino):
    """Pruebas con pycel: el encadenado del saldo, el IVA trimestral, el
    ROUNDUP del umbral y el IVA del CAPEX."""
    import contextlib
    import os
    import shutil

    import openpyxl
    from pycel import ExcelCompiler

    fuera = {'fallos': [], 'grupo_b': {}}

    def _ev(xl, ref):
        with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
            try:
                return xl.evaluate(ref)
            except Exception as e:                           # noqa: BLE001
                return 'ERR:' + type(e).__name__

    os.makedirs(destino, exist_ok=True)

    # --- 03: caso trazado de la SPEC. Saldo 15.000, cobros 40.000, pagos
    #         35.000 en enero → febrero abre en 20.000.
    f03 = '03-cash-flow-forecast.xlsx'
    p03 = os.path.join(carpeta, f03)
    if os.path.isfile(p03):
        copia = os.path.join(destino, '_gb_' + f03)
        shutil.copy2(p03, copia)
        xl = ExcelCompiler(filename=copia)
        _ev(xl, "'Flujo Mensual'!C5")
        _ev(xl, "'Flujo Mensual'!B" + str(F_SALDO_FIN))
        _ev(xl, "'Flujo Mensual'!B" + str(F_TARJETA))
        xl.set_value("'Flujo Mensual'!B5", 15000)
        xl.set_value("'Flujo Mensual'!B7", 40000)
        # Sin desfase de tarjeta ni de proveedores, para aislar el encadenado.
        xl.set_value("'Parámetros'!C4", 0)
        xl.set_value("'Parámetros'!C6", 0)
        for coord in ('B8', 'B9', 'B10', 'B11'):
            xl.set_value("'Flujo Mensual'!" + coord, 0)
        xl.set_value("'Flujo Mensual'!B17", 35000)
        saldo_ene = _ev(xl, "'Flujo Mensual'!B" + str(F_SALDO_FIN))
        apertura_feb = _ev(xl, "'Flujo Mensual'!C5")

        wb = openpyxl.load_workbook(copia)
        formula_c5 = wb['Flujo Mensual']['C5'].value
        fila_saldo = None
        for r in range(1, wb['Flujo Mensual'].max_row + 1):
            if wb['Flujo Mensual'].cell(row=r, column=1).value == 'SALDO FINAL':
                fila_saldo = r
        apunta = (isinstance(formula_c5, str)
                  and formula_c5.replace('$', '').upper()
                  == '=B' + str(fila_saldo))

        # IVA trimestral: abril liquida el 1T.
        xl2 = ExcelCompiler(filename=copia)
        _ev(xl2, "'Flujo Mensual'!E25")
        _ev(xl2, "'Flujo Mensual'!B" + str(F_IVA_REP))
        xl2.set_value("'Flujo Mensual'!B7", 33000)
        iva_abril = _ev(xl2, "'Flujo Mensual'!E25")
        iva_mayo = _ev(xl2, "'Flujo Mensual'!F25")

        fuera['grupo_b']['03'] = {
            'formula_C5': formula_c5, 'fila_SALDO_FINAL': fila_saldo,
            'C5_apunta_al_saldo_final': apunta,
            'saldo_final_enero': saldo_ene, 'apertura_febrero': apertura_feb,
            'esperado': 20000,
            'IVA_liquidado_en_abril_con_33000_de_venta_en_enero': iva_abril,
            'IVA_en_mayo_(no_toca)': iva_mayo,
        }
        if not apunta:
            fuera['fallos'].append(
                '03: C5 no apunta al SALDO FINAL (' + str(formula_c5) + ')')
        if apertura_feb != saldo_ene or apertura_feb != 20000:
            fuera['fallos'].append(
                '03: febrero no abre en 20.000 (' + str(apertura_feb) + ')')
        if not (_es_num(iva_abril) and iva_abril > 0):
            fuera['fallos'].append(
                '03: la liquidación de IVA de abril no calcula ('
                + str(iva_abril) + ')')

    # --- 02: ROUNDUP y break-even de caja.
    f02 = '02-calculadora-punto-equilibrio.xlsx'
    p02 = os.path.join(carpeta, f02)
    if os.path.isfile(p02):
        copia = os.path.join(destino, '_gb_' + f02)
        shutil.copy2(p02, copia)
        xl = ExcelCompiler(filename=copia)
        be = _ev(xl, "'Break-Even'!C8")
        cub = _ev(xl, "'Break-Even'!C10")
        caja = _ev(xl, "'Break-Even'!C16")
        ticket_nec = _ev(xl, "'Break-Even'!C13")
        esc_fijos = _ev(xl, "'Escenarios'!C7")
        esc_fact = _ev(xl, "'Escenarios'!C10")
        wb = openpyxl.load_workbook(copia)
        wse = wb['Escenarios']
        editables = sum(1 for fila in wse['C4:E6'] for c in fila
                        if not c.protection.locked)
        fuera['grupo_b']['02'] = {
            'break_even_mes': be, 'cubiertos_dia_redondeados': cub,
            'break_even_de_caja': caja, 'ticket_medio_necesario': ticket_nec,
            'Escenarios_C7_lee_de_Datos': esc_fijos,
            'Escenarios_C10_con_dias_de_Datos': esc_fact,
            'celdas_editables_en_Escenarios_C4_E6': editables,
        }
        if _es_num(cub) and cub != int(cub):
            fuera['fallos'].append(
                '02: los cubiertos/día no están redondeados (' + str(cub)
                + ')')
        if _es_num(cub) and _es_num(be):
            # Con el umbral redondeado hacia ARRIBA hay que llegar o pasarse.
            ticket = 22.0
            dias = 26
            if cub * ticket * dias < be - 0.01:
                fuera['fallos'].append(
                    '02: con ' + str(cub) + ' cubiertos no se llega al '
                    'break-even de ' + str(be))
        if editables != 9:
            fuera['fallos'].append(
                '02: Escenarios!C4:E6 tiene ' + str(editables)
                + '/9 celdas editables (RD-06)')
        if not (_es_num(caja) and _es_num(be)
                and caja > be):
            fuera['fallos'].append(
                '02: el break-even de caja no supera al operativo ('
                + str(caja) + ' vs ' + str(be) + ')')

    # --- 04: el IVA del CAPEX y la dotación anual.
    f04 = '04-presupuesto-inversion-capex.xlsx'
    p04 = os.path.join(carpeta, f04)
    if os.path.isfile(p04):
        copia = os.path.join(destino, '_gb_' + f04)
        shutil.copy2(p04, copia)
        xl = ExcelCompiler(filename=copia)
        _ev(xl, "'Obra'!E5")
        _ev(xl, "'Obra'!I5")
        _ev(xl, "'Resumen'!D10")
        xl.set_value("'Obra'!B5", 100000)
        total_con_iva = _ev(xl, "'Obra'!E5")
        dotacion = _ev(xl, "'Obra'!I5")
        resumen_iva = _ev(xl, "'Resumen'!D10")
        fuera['grupo_b']['04'] = {
            'base_100000_total_con_IVA': total_con_iva,
            'dotacion_anual_obra_3pct': dotacion,
            'Resumen_D10_total_con_IVA': resumen_iva,
        }
        if not (_es_num(total_con_iva)
                and abs(total_con_iva - 121000) < 0.01):
            fuera['fallos'].append(
                '04: 100.000 € de base no dan 121.000 € con IVA ('
                + str(total_con_iva) + ')')
        if not (_es_num(dotacion) and abs(dotacion - 3000) < 0.01):
            fuera['fallos'].append(
                '04: la dotación anual de obra al 3 % no da 3.000 € ('
                + str(dotacion) + ')')
    return fuera
