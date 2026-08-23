#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
grupo_b.py — §3 de `kit-inventario-v2-SPEC.md`: compras, recepción y mermas
(`03-pedidos-compra.xlsx`, `04-recepcion-mercancias.xlsx`,
`05-control-mermas.xlsx`).

Qué resuelve, por fichero (ids del R1 `auditorias/kit-inventario-R1.json`):

  · 03 — DOM-07/DOM-18/DOM-19, TEC-13/TEC-19/TEC-20, COM-16/COM-17/COM-30.
    El IVA venía escrito a mano al 10 % en las 20 líneas del pedido, así que
    un pedido de barra salía con 11 puntos de menos y un pedido de fruta con
    6 de más — y ese total es el documento que se envía al proveedor. Ahora
    la categoría manda: `H` lo trae con `VLOOKUP` contra `Listas` (que escribe
    el motor, §1.3) y encima lleva el desplegable 4/10/21 para sobrescribirlo.
    Se añaden el desglose por tipo (base, cuota y total por 4/10/21), el
    bloque de emisor —sin él el proveedor no sabe quién pide ni dónde
    entregar—, la hoja `Proveedores` VISIBLE (§7-bis.2) que alimenta el
    desplegable y el `VLOOKUP` de teléfono y pedido mínimo, y un `Historial
    Pedidos` real cuya primera fila está enlazada al pedido en curso.

  · 04 — DOM-03/DOM-04/DOM-05/DOM-20/DOM-21, TEC-04/TEC-17, COM-04/COM-06/
    COM-08. Era el único libro del kit con CERO fórmulas mientras sus propias
    Instrucciones prometían que «las temperaturas fuera de rango se marcan
    automáticamente en rojo». Y su tabla de referencia daba 7 °C para toda la
    «carne fresca», cuando el Reg. (CE) 853/2004 sólo permite 7 °C a canales y
    despiece: la carne picada son 2 °C. Con la tabla de 7 filas, una picada a
    6 °C se aceptaba. Ahora la tabla es NUMÉRICA y legal (13 familias con su
    base normativa), `Control Recepción` tiene el desplegable de familia, el
    `VLOOKUP` del umbral y la columna «Conforme (Tª)» calculada, y el rojo lo
    pone el formato condicional del motor (§1.2) sobre `K` y `M`.

  · 05 — DOM-11/DOM-23, TEC-09, COM-10/COM-12. El registro sí calculaba el
    coste por línea, pero las dos hojas que lo convierten en información
    estaban vacías; y el objetivo estrella del producto —«mermas por debajo
    del 3 % sobre compras»— era INCALCULABLE porque no había ninguna celda
    donde escribir las compras del mes. Ahora `Análisis por Categoría` agrega
    con `COUNTIF`/`SUMIF` sobre las 10 categorías canónicas, el `Dashboard
    Mermas` tiene sus entradas verdes (`B10` compras, `B11` objetivo) y el
    `Plan de Acción` deja de ser un título suelto: nueve columnas y cinco
    causas típicas de merma ya escritas.

CONVENCIONES QUE ESTE MÓDULO RESPETA (§1 y familia):
  * verdes `E8F5E9` = lo escribe el cliente; las calculadas, sin relleno.
    Quien pinta el verde dentro de los bloques de datos es `motor.aplicar_verde`
    (regla: columna sin ninguna fórmula). Aquí sólo se marca a mano lo que cae
    FUERA de esos bloques (cabecera del pedido, emisor, entradas del dashboard).
  * `IFERROR` y doble guarda en toda división y todo producto (§1.8).
  * parámetros en celda, nunca literales dentro de la fórmula: el 3 % del
    objetivo de mermas vive en `05!'Dashboard Mermas'!B11`, no en la fórmula.
  * pycel NO implementa `COUNTA` ni `MODE`: `COUNTIF(rango,"<>")` e
    `INDEX/MATCH` sobre una columna de `COUNTIF` (medido y verificado aquí).

IDEMPOTENCIA — cómo está conseguida:
  * las hojas que no traen datos del cliente se BORRAN y se vuelven a crear
    enteras en cada pasada (`_nueva_hoja`), conservando su posición: la 2.ª
    pasada no puede acumular nada porque no queda nada de la 1.ª.
  * las dos que sí tienen fórmulas de la v1.1 (`03!'Pedido Actual'` y
    `05!'Registro Diario Mermas'`) se tocan en su sitio: la columna se inserta
    una sola vez (centinela: `C8` ya dice «Categoría») y las filas se añaden
    con `motor.expandir_filas`, que tiene su propio centinela; después TODA
    fórmula se reescribe con valor absoluto.
  * mis validaciones se marcan con `promptTitle` = «grupo_b · …» y se limpian
    por esa marca antes de reescribirlas. OJO: no puede empezar por
    `motor.MARCA_DV` («kitinv-v2»), o `motor._limpiar_dv` las borraría en
    `cerrar()` y el fichero saldría sin desplegables.

Python 3.7 / openpyxl 3.1.3: sin walrus, sin f-strings de depuración.
"""
import contextlib
import copy
import datetime
import logging
import os
import shutil

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import motor

logging.disable(logging.CRITICAL)

F03 = '03-pedidos-compra.xlsx'
F04 = '04-recepcion-mercancias.xlsx'
F05 = '05-control-mermas.xlsx'

FICHEROS = [F03, F04, F05]

#: marca de MIS validaciones. No puede empezar por `motor.MARCA_DV`.
MARCA = 'grupo_b'

# ==========================================================================
# `03!'Historial Pedidos'` sale del verde automático del motor
# ==========================================================================
# La regla del motor es «una columna es editable si NINGUNA de sus celdas
# lleva fórmula». En el historial eso no vale: la SPEC pide que la PRIMERA
# fila esté enlazada al pedido en curso (`E4='Pedido Actual'!$H$40`…), y esa
# única fila convertiría en «calculadas» —y por tanto BLOQUEADAS— las columnas
# Fecha, Nº Pedido, Proveedor, Productos, Subtotal, IVA y Total de las otras
# 39 filas. El cliente no podría escribir ni un pedido en su propio historial.
# `SIN_VERDE_AUTO` es el gancho que el motor documenta para exactamente esto
# («los grupos marcan ahí las celdas concretas con `marcar_verde()`»).
motor.SIN_VERDE_AUTO = frozenset(set(motor.SIN_VERDE_AUTO) | {'Historial Pedidos'})

# ==========================================================================
# «IVA %» es un TIPO (10), no un porcentaje (0,10)
# ==========================================================================
# `motor.aplicar_formatos` deduce el formato del texto de la cabecera y la
# primera regla de su tabla es «si pone % → 0,0%». La columna del IVA del
# pedido guarda el TIPO —10, 21, 4—, no la fracción, porque así lo pide la
# SPEC (`$H9/100` en el total y desplegable «4,10,21»). Con el formato de
# porcentaje, ese 10 se imprimía «1000,0 %» y el desplegable parecía roto.
# Medido en la 1.ª pasada de este módulo: `03!'Pedido Actual'!H9` con
# `nf=0.0%` y cache 10.
# Se antepone una regla MÁS ESPECÍFICA a la tabla del motor (se para en la
# primera coincidencia). No afecta a ninguna otra cabecera del kit: la única
# que contiene «iva %» es ésta — el historial dice «IVA (€)», que sigue
# cayendo en la regla del euro.
motor.FORMATO_POR_CABECERA = ([(motor.FMT_ENT, ('iva %',))]
                              + list(motor.FORMATO_POR_CABECERA))


# ==========================================================================
# 03 · Datos precargados
# ==========================================================================
#: Proveedores de ejemplo: uno por familia de compra, con las mismas 10
#: categorías canónicas del kit (§1.1). Teléfonos y correos son PLACEHOLDERS
#: (`600 000 00x`, `@ejemplo.es`): un dato de contacto verosímil pero
#: inventado acabaría siendo el de un negocio real.
#: A Proveedor · B Categoría · C Teléfono · D Email · E Pedido mínimo (€) ·
#: F Día de pedido · G Plazo (días) · H Notas
_D = datetime.date

#: Ronda 2 · RD-01/RC-07 — el kit enviaba DOS directorios ficticios: el 03 y
#: el 04 hablaban de «Cárnicas del Mercado» y el 02, el 07 y el BONUS-09 de
#: «Cárnicas del Norte». Un jefe de compras no puede cotejar el albarán de uno
#: contra la ficha homologada del otro. Ahora los nombres salen de
#: `motor.PROVEEDORES_EJEMPLO` y los datos comerciales son LOS MISMOS que los
#: del directorio del 02 (RD-20): teléfono, correo, pedido mínimo, día de
#: pedido y plazo de entrega copiados de `grupo_a.PROVEEDORES`.
#: A Proveedor · B Categoría · C Teléfono · D Email · E Pedido mínimo (€) ·
#: F Día de pedido · G Plazo (días) · H Notas
_DATOS_PROVEEDOR = [
    ('948 21 34 55', 'pedidos@ejemplo-carnicas.es', 150,
     'Lun y jue antes de 12:00', 1,
     'Exigir el albarán con el nº de lote y la temperatura de llegada'),
    ('986 44 12 09', 'pedidos@ejemplo-pescados.es', 120,
     'Diario antes de 17:00', 1,
     'Llega en hielo fundente; rechazar por encima de 2 °C'),
    ('963 55 71 20', 'pedidos@ejemplo-huerta.es', 80,
     'Lun, mié y vie antes de 18:00', 1,
     'Producto de temporada: confirma el precio cada semana'),
    ('954 33 90 41', 'pedidos@ejemplo-economato.es', 250,
     'Mar antes de 14:00', 2,
     'Rappel del 2 % a partir de 2.000 € al mes'),
    ('965 12 78 30', 'pedidos@ejemplo-bebidas.es', 200,
     'Lun antes de 12:00', 2,
     'Cesión de barriles y CO2: revisar el depósito de envases'),
    ('916 40 22 18', 'pedidos@ejemplo-higiene.es', 100,
     'Mié antes de 16:00', 2,
     'Pedir la ficha de datos de seguridad de cada producto'),
]

PROVEEDORES = [
    (motor.PROVEEDORES_EJEMPLO[i], motor.CATEGORIA_PROVEEDOR[i],
     d[0], d[1], d[2], d[3], d[4], d[5])
    for i, d in enumerate(_DATOS_PROVEEDOR)
]

#: Líneas de ejemplo del pedido.
#:
#: Ronda 2 · RT-09 — iban a propósito sin cantidad ni precio «para que no se
#: colaran en un pedido real», y el resultado era que el fichero estrella del
#: bloque de compras se abría con Subtotal y Total en blanco, el bloque de
#: totales a 0,00 € y el historial anunciando «3 líneas» por «0,00 €». Es el
#: ÚNICO libro del kit que abría enseñando ceros, y justo el que el cliente
#: envía al proveedor. Llevan cantidad y precio —los del 01— y el aviso de
#: que hay que borrarlas está en la nota de la hoja y en las Instrucciones.
#: Los tres productos cubren los tres tipos de IVA: 10 % la carne, 21 % el
#: vino y 4 % la lechuga (DOM-07), así que el desglose por tipo se ve entero.
#: (producto del 01, categoría, unidad, cantidad, precio/ud del 01)
LINEAS_PEDIDO = [
    ('Solomillo de ternera', 'Cárnicos', 'kg', 5, 32.00),
    ('Vino tinto (botella 75 cl)', 'Bebidas Alcohólicas', 'ud', 12, 4.80),
    ('Lechuga', 'Verduras/Frutas', 'ud', 10, 0.95),
]

ESTADOS_PEDIDO = ['Borrador', 'Enviado', 'Confirmado', 'Recibido',
                  'Recibido con incidencia', 'Facturado']

#: Historial de ejemplo.
#:
#: Ronda 2 · RD-28/RT-20 — las fechas iban con `TODAY()-n`, así que el
#: historial se recolocaba solo cada vez que se abría el libro: un historial
#: cuyo contenido cambia solo es lo contrario de un historial, contamina el
#: cotejo con el albarán (que se hace por fecha) y además metía una FÓRMULA
#: VOLÁTIL dentro de tres celdas VERDES, que por convención de la familia son
#: dato del cliente. Ahora son constantes, como las del 04 y las del 05.
#: Las cuotas son coherentes con el tipo de la categoría del proveedor: 10 %
#: la carne, 21 % la bodega, 4 % la fruta.
HISTORIAL = [
    (_D(2026, 7, 20), 'PED-2026-001', motor.PROVEEDORES_MARCADOS[0],
     '6 líneas', 342.80, 34.28, 377.08, 'Facturado', 'Sin incidencias'),
    (_D(2026, 8, 3), 'PED-2026-002', motor.PROVEEDORES_MARCADOS[4],
     '4 líneas', 218.50, 45.89, 264.39, 'Recibido',
     'Barril de 30 L en depósito'),
    (_D(2026, 8, 14), 'PED-2026-003', motor.PROVEEDORES_MARCADOS[2],
     '9 líneas', 96.40, 3.86, 100.26, 'Recibido con incidencia',
     'Dos cajas de tomate retiradas'),
]

# ==========================================================================
# 04 · Tabla LEGAL de temperaturas de recepción (DOM-05/COM-08)
# ==========================================================================
# Familia · Tª ideal (texto) · Tª máx. aceptable (NÚMERO, o «N/A») · norma.
# La columna C es la que lee el `VLOOKUP` del registro: por eso es un número
# y no «7°C máx.» como en la v1.1, donde el criterio existía en papel y no se
# podía aplicar a ningún dato.
#
# Las 5 primeras filas desdoblan lo que la v1.1 llamaba «Carne fresca ≤ 7 °C»:
# 7 °C sólo vale para canales y despiece de ungulados. Con la tabla vieja, una
# carne picada a 6 °C —el triple de su límite— se daba por buena.
TEMPERATURAS = [
    ('Canal y despiece de ungulados domésticos', '0 a 7 °C', 0, 7,
     'Reg. (CE) 853/2004, Anexo III, Secc. I, Cap. VII'),
    ('Despojos comestibles', '0 a 3 °C', 0, 3,
     'Reg. (CE) 853/2004, Anexo III, Secc. I, Cap. VII'),
    ('Aves y lagomorfos', '0 a 4 °C', 0, 4,
     'Reg. (CE) 853/2004, Anexo III, Secc. II, Cap. V'),
    ('Preparados de carne', '0 a 4 °C', 0, 4,
     'Reg. (CE) 853/2004, Anexo III, Secc. V, Cap. III'),
    ('Carne picada', '0 a 2 °C', 0, 2,
     'Reg. (CE) 853/2004, Anexo III, Secc. V, Cap. III'),
    ('Pescado fresco y marisco', '0 a 2 °C, en hielo fundente', 0, 2,
     'Reg. (CE) 853/2004, Anexo III, Secc. VIII, Cap. VII'),
    # RT-14 · el libro se contradecía a sí mismo en la única familia donde la
    # norma admite tolerancia: la nota citaba los -15 °C de transporte y el
    # umbral computable era -18, así que un congelado a -16 —dentro de la
    # tolerancia que el propio fichero cita— salía «RECHAZAR» y se pintaba de
    # rojo. -18 °C es la temperatura de CONSERVACIÓN; la de RECEPCIÓN es -15.
    # El mínimo (-40) no es un capricho: por debajo de -40 lo que falla es el
    # termómetro, no el camión.
    ('Congelados',
     '-18 °C en conservación; hasta -15 °C admisible al recibir', -40, -15,
     'Reg. (CE) 853/2004; la tolerancia breve de -15 °C en transporte es la '
     'que se aplica AL RECIBIR. Mantener a -18 °C una vez almacenado.'),
    ('Comidas preparadas refrigeradas (más de 24 h)', '0 a 4 °C', 0, 4,
     'RD 3484/2000, art. 6'),
    ('Comidas preparadas refrigeradas (menos de 24 h)', '0 a 8 °C', 0, 8,
     'RD 3484/2000, art. 6'),
    ('Lácteos y otros refrigerados', '0 a 8 °C o lo que diga la etiqueta',
     0, 8, 'Reg. (CE) 853/2004 y etiquetado del fabricante'),
    ('Frutas y verduras', '4 a 12 °C', 0, 12,
     'Buenas prácticas; sin límite legal de recepción'),
    ('Huevos', 'Temperatura ambiente constante, sin cambios bruscos; '
     'no refrigerar antes de la venta', 'N/A', 'N/A',
     'Reg. (CE) 589/2008, art. 2'),
    ('Secos y economato (ambiente)', 'Lugar fresco y seco, menos de 25 °C',
     'N/A', 'N/A', 'Reg. (CE) 852/2004, Anexo II, Cap. IX'),
    # RD-23 · las dos familias que faltaban. Sin ellas, un barril de cerveza,
    # una caja de guantes o un paquete de servilletas —que el kit precarga en
    # el 01— había que archivarlos bajo «Secos y economato» o dejar la
    # columna en blanco, y sin familia no hay veredicto ni registro.
    ('Bebidas y conservas (ambiente)',
     'Lugar fresco y seco; la cerveza de barril, entre 5 y 12 °C',
     'N/A', 'N/A',
     'Sin límite legal de recepción; sigue la indicación del fabricante'),
    ('No alimentario (limpieza, menaje y desechables)',
     'Ambiente, en almacén separado de los alimentos', 'N/A', 'N/A',
     'Reg. (CE) 852/2004, Anexo II, Cap. IX: los productos químicos van '
     'separados y señalizados, nunca junto a alimentos'),
]

#: RD-23 · el PUENTE entre los dos vocabularios del kit: las 10 categorías
#: comerciales (que llevan los otros ocho ficheros) y las familias NORMATIVAS
#: de la tabla de arriba. Separarlas es correcto —el umbral legal lo fija la
#: norma, no tu categoría de compra— pero el usuario se quedaba solo ante el
#: mapeo. Ahora la columna «Familia sugerida» del registro la propone sola con
#: un VLOOKUP contra esta tabla y el usuario sólo la afina cuando hace falta
#: (carne picada, despojos, comida preparada).
PUENTE_CATEGORIA_FAMILIA = [
    ('Cárnicos', 'Canal y despiece de ungulados domésticos'),
    ('Pescados', 'Pescado fresco y marisco'),
    ('Lácteos', 'Lácteos y otros refrigerados'),
    ('Verduras/Frutas', 'Frutas y verduras'),
    ('Secos/Granos', 'Secos y economato (ambiente)'),
    ('Congelados', 'Congelados'),
    ('Bebidas Alcohólicas', 'Bebidas y conservas (ambiente)'),
    ('Bebidas No Alcohólicas', 'Bebidas y conservas (ambiente)'),
    ('Limpieza', 'No alimentario (limpieza, menaje y desechables)'),
    ('Otros', 'Secos y economato (ambiente)'),
]

#: Última fila de la tabla legal y del puente (para los VLOOKUP y la DV).
R1_TEMP = 3 + len(TEMPERATURAS)
R1_PUENTE = 3 + len(PUENTE_CATEGORIA_FAMILIA)

#: Fechas de los ejemplos del 04. Van escritas (no `TODAY()`) porque sus
#: columnas son VERDES y una fórmula ahí convertiría toda la columna en
#: «calculada» para `motor.aplicar_verde` — y el cliente no podría escribir la
#: fecha de sus propias entregas. Se eligen dentro del mes de la versión.
#:
#: Ronda 2 · RD-03 — las dos hojas de ejemplo se CONTRADECÍAN sobre los mismos
#: albaranes: la incidencia reclamaba 64 € por 2 kg que faltaban en un albarán
#: que la recepción registraba como servido completo y conforme, y la falta
#: real que sí registraba la recepción no aparecía en el registro de
#: incidencias. El fichero que se vende como prueba documental ante una
#: inspección enseñaba a llevar dos registros que no cuadran. Ahora los
#: ejemplos son UNA historia: mismo albarán, misma cantidad, misma familia, y
#: el importe reclamado ES la diferencia por el precio.
#: (fecha, proveedor, albarán, producto, categoría del kit, familia, lote,
#:  pedido, recibido, precio/ud, caducidad, temp, calidad, incidencia,
#:  receptor)
EJEMPLOS_RECEPCION = [
    (_D(2026, 8, 18), motor.PROVEEDORES_MARCADOS[0], 'ALB-24518',
     'Solomillo de ternera (ejemplo)', 'Cárnicos',
     'Canal y despiece de ungulados domésticos', 'L-260817', 12, 10, 32.00,
     _D(2026, 8, 25), 3, '✓',
     'Faltan 2 kg de los 12 pedidos: reclamado por correo el mismo día',
     'Jefe de cocina'),
    (_D(2026, 8, 18), motor.PROVEEDORES_MARCADOS[1], 'ALB-0091',
     'Merluza fresca (ejemplo)', 'Pescados', 'Pescado fresco y marisco',
     'L-260818', 8, 8, 14.00, _D(2026, 8, 21), 5, '✗',
     'Llega a 5 °C con el límite de su familia en 2 °C: partida completa '
     'devuelta al transportista',
     'Segundo de cocina'),
    (_D(2026, 8, 19), motor.PROVEEDORES_MARCADOS[2], 'ALB-7742',
     'Tomate (ejemplo)', 'Verduras/Frutas', 'Frutas y verduras', 'L-260819',
     30, 30, 1.95, _D(2026, 8, 26), 9, '✗',
     'Dos cajas de 10 kg con moho en el fondo: retiradas y anotadas en el '
     'control de mermas',
     'Encargado de compras'),
    # La cuarta línea existe para enseñar el «N/A»: los huevos no se
    # controlan por temperatura en la recepción (RD-04/RD-23).
    (_D(2026, 8, 19), motor.PROVEEDORES_MARCADOS[2], 'ALB-7743',
     'Huevos M, docena (ejemplo)', 'Otros', 'Huevos', 'L-260819', 20, 20,
     2.20, _D(2026, 9, 5), 17, '✓', '', 'Encargado de compras'),
]

TIPOS_INCIDENCIA = ['Cantidad incorrecta', 'Temperatura fuera de rango',
                    'Producto en mal estado', 'Caducidad demasiado corta',
                    'Producto no solicitado', 'Precio distinto al pactado',
                    'Falta el lote o el albarán']

ESTADOS_INCIDENCIA = ['Abierta', 'Reclamada', 'Abono recibido',
                      'Cerrada sin abono', 'Cerrada']

#: RD-03 · cada incidencia CASA con su línea de recepción: mismo albarán,
#: mismo proveedor, mismo producto y un importe que sale de la diferencia (o
#: de la cantidad devuelta) por el precio de la propia recepción.
EJEMPLOS_INCIDENCIA = [
    (_D(2026, 8, 18), 'ALB-24518', motor.PROVEEDORES_MARCADOS[0],
     'Solomillo de ternera (ejemplo)', 'Cantidad incorrecta',
     'Faltan 2 kg de los 12 pedidos: sólo se reciben 10',
     'Reclamado por correo el mismo día', 'Jefe de cocina',
     64.00, 0, 'Reclamada'),          # 2 kg x 32,00 €/kg
    (_D(2026, 8, 18), 'ALB-0091', motor.PROVEEDORES_MARCADOS[1],
     'Merluza fresca (ejemplo)', 'Temperatura fuera de rango',
     'Llega a 5 °C con el límite de su familia en 2 °C',
     'Partida completa (8 kg) devuelta al transportista en el momento',
     'Segundo de cocina', 112.00, 112.00, 'Abono recibido'),   # 8 x 14,00
    (_D(2026, 8, 19), 'ALB-7742', motor.PROVEEDORES_MARCADOS[2],
     'Tomate (ejemplo)', 'Producto en mal estado',
     'Dos cajas de 10 kg con moho en el fondo',
     'Retiradas y anotadas en el control de mermas',
     'Encargado de compras', 39.00, 0, 'Abierta'),             # 20 x 1,95
]

# ==========================================================================
# 05 · Mermas
# ==========================================================================
#: Motivos reales de merma en cocina. Son el vocabulario del desplegable de
#: `Registro Diario Mermas!H` Y las etiquetas del bloque auxiliar de
#: `Análisis por Categoría!H4:I10`, sobre el que corre el INDEX/MATCH del
#: «motivo más frecuente» (pycel no implementa `MODE`, §1).
#: Son DIEZ, tantos como filas verdes tiene el bloque auxiliar, para que
#: ningún hueco quede sin su contador. Y NINGUNO lleva coma dentro: una lista
#: inline de Excel separa por comas, así que «Rotura, derrame o caída» se
#: partía en dos opciones —«Rotura» y «derrame o caída»— y ninguna de las dos
#: casaba con el COUNTIF. Lo mismo vale para cualquier lista de `_lista_inline`.
MOTIVOS = [
    'Caducidad superada',
    'Mal estado en recepción',
    'Rotura o derrame',
    'Error de elaboración',
    'Sobreproducción',
    'Devolución de cliente',
    'Cadena de frío rota',
    'Merma de despiece',
    'Plato cambiado o rechazado',
    'Descuadre sin justificar',
]

#: Tres líneas de ejemplo. SIN fecha a propósito: la columna A es verde y una
#: fórmula `TODAY()` la volvería «calculada» para el motor. Con ellas, el
#: análisis y el dashboard arrancan con números de verdad en vez de ceros.
EJEMPLOS_MERMA = [
    ('Solomillo de ternera (ejemplo)', 'Cárnicos', 1.2, 'kg', 32.00,
     'Error de elaboración', 'Jefe de cocina',
     'Ficha de despiece con rendimiento objetivo y pesada del recorte'),
    ('Merluza fresca (ejemplo)', 'Pescados', 2.5, 'kg', 14.00,
     'Cadena de frío rota', 'Segundo de cocina',
     'Revisar la cámara dos veces al día y el cierre de la puerta'),
    # el precio es el del 01 (0,95 €/ud): todo producto que aparece en dos
    # ficheros del kit tiene que valer lo mismo en los dos (RD-11).
    ('Lechuga (ejemplo)', 'Verduras/Frutas', 6, 'ud', 0.95,
     'Caducidad superada', 'Encargado de compras',
     'Etiquetar la fecha de apertura y servir por orden de caducidad'),
]

PRIORIDADES = ['Alta', 'Media', 'Baja']
ESTADOS_PLAN = ['Pendiente', 'En curso', 'Implantada', 'Verificada',
                'Descartada']

#: Las cinco causas que pide la SPEC, escritas como las escribiría un jefe de
#: cocina. El «coste mensual evitado» es una ESTIMACIÓN del propio cliente
#: (celda editable), no una promesa del producto.
PLAN_ACCION = [
    ('Alta', 'Cárnicos',
     'Se tira producto elaborado al cierre del servicio',
     'Sobreproducción: se produce por costumbre, no contra la previsión de '
     'cubiertos',
     'Producir contra la previsión del día y anotar el sobrante en el '
     'registro de mermas antes de cerrar',
     'Jefe de cocina', 180, 'Pendiente'),
    ('Alta', 'Verduras/Frutas',
     'Aparece género caducado con existencias más nuevas delante',
     'El etiquetado FEFO está incompleto: no se anota la fecha de apertura',
     'Etiquetar toda apertura con la fecha y colocar delante lo que antes '
     'caduca',
     'Segundo de cocina', 120, 'En curso'),
    ('Media', 'Pescados',
     'Género perdido por rotura de la cadena de frío',
     'La cámara no se revisa a diario y la puerta queda abierta en servicio',
     'Registro de temperaturas dos veces al día y revisión del cierre de '
     'puerta al terminar el servicio',
     'Responsable del plan APPCC', 95, 'Pendiente'),
    ('Media', 'Cárnicos',
     'La merma de despiece se va por encima de lo previsto',
     'Cada cocinero limpia distinto: no hay ficha de despiece',
     'Ficha de despiece con rendimiento objetivo y pesada del recorte una '
     'vez por semana',
     'Jefe de partida', 75, 'Pendiente'),
    ('Baja', 'Secos/Granos',
     'Existencias muertas en el economato',
     'Se pide por encima del consumo real: no se mira el stock antes de '
     'pedir',
     'Pedir contra la columna A Pedir del inventario y respetar el punto de '
     'pedido',
     'Encargado de compras', 40, 'Pendiente'),
]


# ==========================================================================
# Utilidades
# ==========================================================================
def _reg(registro, ws, coord, formula):
    registro.append((ws.title, coord, formula))
    return formula


def _f(ws, coord, formula, registro, fmt=None):
    """Escribe una FÓRMULA y la anota para que `main.py` verifique su cache."""
    cel = ws[coord]
    cel.value = formula
    if fmt:
        cel.number_format = fmt
    _reg(registro, ws, coord, formula)
    return cel


def _verde(cel, fmt=None):
    cel.fill = PatternFill('solid', fgColor=motor.VERDE)
    cel.protection = Protection(locked=False)
    if fmt:
        cel.number_format = fmt
    return cel


def _nueva_hoja(wb, nombre, despues=None):
    """Hoja desde CERO conservando su posición.

    Borrar y recrear (en vez de vaciar celda a celda) es lo que garantiza la
    idempotencia: no quedan estilos, combinaciones, validaciones ni formatos
    condicionales de la pasada anterior que pudieran acumularse.
    """
    idx = None
    if nombre in wb.sheetnames:
        idx = wb.sheetnames.index(nombre)
        del wb[nombre]
    elif despues and despues in wb.sheetnames:
        idx = wb.sheetnames.index(despues) + 1
    if idx is None:
        return wb.create_sheet(nombre)
    return wb.create_sheet(nombre, idx)


def _titulo(ws, fila, texto, col_fin):
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font = Font(bold=True, size=13)
    cel.alignment = Alignment(vertical='center')
    _merge(ws, 'A{f}:{c}{f}'.format(f=fila, c=get_column_letter(col_fin)))
    return cel


def _nota(ws, fila, texto, col_fin, col_ini=1):
    cel = ws.cell(row=fila, column=col_ini, value=texto)
    cel.font = Font(size=9, italic=True)
    cel.alignment = Alignment(vertical='top', wrap_text=True)
    _merge(ws, '{a}{f}:{b}{f}'.format(a=get_column_letter(col_ini), f=fila,
                                      b=get_column_letter(col_fin)))
    return cel


def _cabecera(ws, fila, textos, color=None):
    color = color or motor.CAB
    for i, t in enumerate(textos):
        cel = ws.cell(row=fila, column=1 + i, value=t)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = PatternFill('solid', fgColor=color)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)


def _anchos(ws, anchos):
    for i, a in enumerate(anchos):
        ws.column_dimensions[get_column_letter(1 + i)].width = a


def _cebra(ws, r0, r1, ncols):
    """Banda gris de las filas pares, como la v1.1. Las columnas editables las
    repinta de verde `motor.aplicar_verde` después."""
    for fila in range(r0, r1 + 1):
        if fila % 2:
            continue
        for c in range(1, ncols + 1):
            ws.cell(row=fila, column=c).fill = PatternFill(
                'solid', fgColor=motor.BANDA)


def _merge(ws, ref):
    """`merge_cells` idempotente: repetir la combinación dejaría dos entradas
    iguales en `merged_cells.ranges` y la 2.ª pasada daría diferencia."""
    if ref not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells(ref)


def _limpiar_dv(ws):
    """Deja SÓLO las validaciones del motor: borra las mías (para volver a
    escribirlas) y las HEREDADAS de la v1.1.

    Lo segundo no es cosmética. `05!'Registro Diario Mermas'!H4:H53` traía una
    lista de motivos vieja («Caducidad, Preparación, Devolución, Accidente…»)
    que NO coincide con las etiquetas sobre las que corre el `COUNTIF` del
    análisis: dejarla puesta significaba dos desplegables solapados en la
    misma columna y un «motivo más frecuente» que nunca contaría nada.
    """
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if (getattr(dv, 'promptTitle', None) or '').startswith(motor.MARCA_DV)]


def _dv(ws, ref, formula1, titulo, prompt):
    dv = DataValidation(
        type='list', formula1=formula1, allow_blank=True,
        showErrorMessage=True, errorTitle=titulo,
        error='Elige un valor de la lista. Si cada línea usa palabras '
              'distintas, los COUNTIF y los SUMIF de los análisis no suman '
              'nada.',
        errorStyle='stop', showInputMessage=True,
        promptTitle='{} · {}'.format(MARCA, titulo), prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(ref)
    return dv


def _lista_inline(valores):
    """Lista inline de DV. Excel corta en 255 caracteres CONTANDO las comillas
    y no avisa: mejor reventar aquí que publicar un desplegable truncado."""
    for v in valores:
        if ',' in v:
            raise ValueError('«{}» lleva una coma: en una lista inline de '
                             'Excel se partiría en dos opciones'.format(v))
    formula = '"{}"'.format(','.join(valores))
    if len(formula) > 255:
        raise ValueError('DV inline de {} caracteres (>255)'.format(
            len(formula)))
    return formula


def _instrucciones(ws, lineas):
    """Reescribe `Instrucciones` entera y deja la línea de versión en una fila
    FIJA.

    `motor.bio_y_version` se ancla en esa línea para escribir debajo la nota
    de desprotección, la bio y la versión. Si la fila de anclaje se moviera
    entre pasadas (p. ej. por dejar que `max_row` creciera), el bloque bajaría
    dos filas en cada pasada y la idempotencia daría diferencias.
    """
    ultima = ws.max_row
    for fila in range(1, max(ultima, len(lineas) + 6) + 1):
        ws.cell(row=fila, column=1).value = None
    for i, texto in enumerate(lineas):
        cel = ws.cell(row=fila_i(i), column=1, value=texto)
        cel.alignment = Alignment(vertical='top', wrap_text=True)
        if i == 0:
            cel.font = Font(bold=True, size=13)
    ancla = fila_i(len(lineas)) + 1
    ws.cell(row=ancla, column=1).value = motor.VERSION_LINE
    ws.column_dimensions['A'].width = 100
    return ancla


def fila_i(i):
    """Fila de la línea `i` de Instrucciones: 1, 2, y luego una en blanco."""
    return 1 + i if i < 2 else 2 + i


# ==========================================================================
# API del grupo
# ==========================================================================
def pre(wb, fname, informe):
    """Lo ÚNICO que va antes del motor: insertar «Categoría» en el 03.

    Tiene que ser aquí porque el motor fija en `cerrar()` la DV de categorías
    sobre `03!C` guiándose por el centinela `C8`; si la columna se insertara
    después, esa DV habría caído sobre la columna de unidades.
    """
    if fname != F03:
        return
    ws = wb['Pedido Actual']
    if ws['C8'].value == 'Categoría':
        return                                        # 2.ª pasada
    motor.insertar_columna(ws, 3)
    ws['C8'] = 'Categoría'
    informe.append('03:Pedido Actual!C: columna «Categoría» insertada; el IVA '
                   'de cada línea sale de ella (DOM-07)')


def post(wb, fname, informe, registro):
    if fname == F03:
        _pedidos(wb, informe, registro)
    elif fname == F04:
        _recepcion(wb, informe, registro)
    elif fname == F05:
        _mermas(wb, informe, registro)


# ==========================================================================
# 03 · Pedidos de compra
# ==========================================================================
ANCHOS_03 = [5, 32, 22, 14, 14, 16, 18, 9, 16]
R0_03, R1_03 = 9, 38             # §1.9: el pedido pasa de 20 a 30 líneas
FILA_BASE, FILA_CUOTA, FILA_TOTAL = 40, 41, 42
FILA_MINIMO = 43                 # RD-17 · veredicto del pedido mínimo
FILA_DESGLOSE = 45               # cabecera; 46-48 = 4 %, 10 %, 21 %


def _pedidos(wb, informe, registro):
    _pedido_actual(wb, informe, registro)
    _proveedores_03(wb, informe, registro)
    _historial_03(wb, informe, registro)
    _instrucciones_03(wb, informe)


def _pedido_actual(wb, informe, registro):
    ws = wb['Pedido Actual']
    _limpiar_dv(ws)
    _anchos(ws, ANCHOS_03)

    # ---- filas libres (§1.9) -------------------------------------------
    # La cola de la v1.1 (SUBTOTAL/TOTAL en 30-31) NO se arrastra: cae dentro
    # del bloque nuevo y la propia replicación la borra. Los totales se
    # reescriben enteros más abajo, en 40-42.
    anadidas = motor.expandir_filas(ws, 28, R1_03, cola=(), numerar=1)
    if anadidas:
        informe.append('03:Pedido Actual: {} líneas nuevas (20 → 30)'
                       .format(anadidas))

    # ---- cabecera del pedido (TEC-19/COM-30) ---------------------------
    # La columna A mide 5 caracteres: con la etiqueta suelta en A3, el
    # documento que recibía el proveedor decía «Prove».
    etiquetas = [(3, 'Proveedor:'), (4, 'Fecha del pedido:'),
                 (5, 'Fecha de entrega:'), (6, 'Nº de pedido:')]
    for fila, texto in etiquetas:
        cel = ws.cell(row=fila, column=1, value=texto)
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal='right', vertical='center')
        _merge(ws, 'A{f}:B{f}'.format(f=fila))
        _verde(ws.cell(row=fila, column=3))
    ws['C4'].number_format = motor.FMT_FECHA
    ws['C5'].number_format = motor.FMT_FECHA

    _dv(ws, 'C3', '=Proveedores!$A$4:$A$23', 'Proveedor no dado de alta',
        'Los proveedores salen de la hoja Proveedores de este mismo libro. '
        'Da de alta ahí a los tuyos y aparecerán en esta lista.')
    # RT-09/RD-17 · la cabecera se siembra para que el libro ABRA funcionando:
    # sin proveedor elegido, el teléfono, el pedido mínimo y el veredicto de
    # la fila 43 salen todos en blanco y la cabecera parece rota.
    ws['C3'] = motor.PROVEEDORES_MARCADOS[0]
    ws['C4'] = _D(2026, 8, 18)
    ws['C5'] = _D(2026, 8, 19)
    ws['C6'] = 'PED-2026-004'
    _nota(ws, 2, 'EJEMPLO: la cabecera y las tres primeras líneas vienen '
                 'rellenas para que veas el pedido funcionando (IVA por '
                 'línea, base, cuota, total, desglose por tipo y aviso de '
                 'pedido mínimo). BÓRRALAS antes de enviar tu primer pedido.',
          9)

    for fila, texto in ((3, 'Teléfono'), (4, 'Pedido mín. (€)')):
        cel = ws.cell(row=fila, column=4, value=texto)
        cel.font = Font(bold=True, size=9)
        cel.alignment = Alignment(horizontal='right', vertical='center')
    _f(ws, 'E3', '=IF($C$3="","",IFERROR(VLOOKUP($C$3,Proveedores!$A$4:$H$23,'
                 '3,FALSE),""))', registro)
    _f(ws, 'E4', '=IF($C$3="","",IFERROR(VLOOKUP($C$3,Proveedores!$A$4:$H$23,'
                 '5,FALSE),""))', registro, motor.FMT_EUR)

    # bloque de EMISOR (COM-30): sin él, quien recibe el pedido no sabe quién
    # pide ni dónde entregar.
    emisor = [(3, 'Establecimiento'), (4, 'CIF / NIF'),
              (5, 'Entrega en'), (6, 'Contacto')]
    for fila, texto in emisor:
        cel = ws.cell(row=fila, column=6, value=texto)
        cel.font = Font(bold=True, size=9)
        cel.alignment = Alignment(horizontal='right', vertical='center')
        _verde(ws.cell(row=fila, column=7))

    # ---- cabecera de la tabla ------------------------------------------
    _cabecera(ws, 8, ['#', 'Producto', 'Categoría', 'Unidad', 'Cantidad',
                      'Precio/ud (€)', 'Subtotal (€)', 'IVA %', 'Total (€)'])

    # ---- las 30 líneas --------------------------------------------------
    for fila in range(R0_03, R1_03 + 1):
        ws.cell(row=fila, column=1).value = fila - R0_03 + 1
        # §1.8 · doble guarda: la de la v1.1 vigilaba sólo la cantidad, así
        # que una línea con cantidad y sin precio valía 0,00 € y se sumaba al
        # total como si el proveedor la regalara.
        _f(ws, 'G{}'.format(fila), motor.guarda_doble('E', 'F', fila),
           registro, motor.FMT_EUR)
        # DOM-07 · el tipo sale de la categoría; nunca de un 10 escrito a
        # mano. Ronda 2 (RD-10/RT-16): dos correcciones.
        #  · Ninguna asignación POR CATEGORÍA puede acertar siempre: en
        #    «Secos/Granos» conviven el arroz al 4 % y el aceite al 10 %, y en
        #    «Bebidas No Alcohólicas» el agua al 10 % y la tónica al 21 %. Se
        #    busca PRIMERO el producto en la tabla de excepciones de `Listas`
        #    (E:F) y sólo se cae a la categoría si no está.
        #  · Si la línea tiene producto y se queda SIN categoría, el IFERROR
        #    dejaba caer el tipo al 10 % en silencio: en una línea de alcohol,
        #    limpieza o desechables son 11 puntos de menos. Ahora la casilla
        #    se queda en blanco, que es lo que se ve.
        _f(ws, 'H{}'.format(fila),
           '=IF($B{f}="","",IFERROR(VLOOKUP($B{f},Listas!$E$2:$F$20,2,FALSE),'
           'IF($C{f}="","",IFERROR(VLOOKUP($C{f},Listas!$A$2:$B$11,2,FALSE),'
           '""))))'.format(f=fila), registro, motor.FMT_ENT)
        _f(ws, 'I{}'.format(fila),
           '=IF(OR($E{f}="",$F{f}=""),"",$E{f}*$F{f}*(1+$H{f}/100))'
           .format(f=fila), registro, motor.FMT_EUR)

    _dv(ws, 'H{}:H{}'.format(R0_03, R1_03), _lista_inline(['4', '10', '21']),
        'Tipo de IVA no válido',
        'El tipo lo trae la categoría, pero puedes sobrescribirlo aquí: pan, '
        'harinas, legumbres y cereales van al 4 %; los refrescos con azúcares '
        'añadidos, al 21 %. Para escribir encima de la fórmula: Revisar → '
        'Desproteger hoja (no tiene contraseña).')

    for i, datos in enumerate(LINEAS_PEDIDO):
        fila = R0_03 + i
        producto, categoria, unidad, cantidad, precio = datos
        # OJO al nombre: va SIN el «(ejemplo)» pegado porque es la clave con
        # la que el IVA busca el producto en `Listas!E:F` (RD-10). El aviso de
        # que son líneas de ejemplo va en la nota de la hoja y en el propio
        # bloque de Instrucciones, que es donde se lee.
        ws.cell(row=fila, column=2).value = producto
        ws.cell(row=fila, column=3).value = categoria
        ws.cell(row=fila, column=4).value = unidad
        ws.cell(row=fila, column=5).value = cantidad
        cel = ws.cell(row=fila, column=6, value=precio)
        cel.number_format = motor.FMT_EUR

    # ---- totales (TEC-20/DOM-18) ---------------------------------------
    # En la v1.1 el «TOTAL CON IVA» se imprimía bajo la cabecera «Subtotal» y
    # no existía la fila de CUOTA, que es justo la que pide el historial.
    # OJO: deshacer las combinaciones ANTES de limpiar. En la 2.ª pasada esas
    # filas ya están combinadas y las celdas de la derecha son `MergedCell`,
    # cuyo `value` es de sólo lectura («object attribute 'value' is
    # read-only»).
    for ref in [str(r) for r in ws.merged_cells.ranges]:
        primera = int(''.join(ch for ch in ref.split(':')[0] if ch.isdigit()))
        if FILA_BASE <= primera <= FILA_TOTAL:
            ws.unmerge_cells(ref)
    for fila in range(FILA_BASE, FILA_TOTAL + 1):
        for col in range(1, 10):
            ws.cell(row=fila, column=col).value = None
    totales = [
        (FILA_BASE, 'BASE IMPONIBLE (€)', '=SUM($G${}:$G${})'.format(
            R0_03, R1_03)),
        (FILA_CUOTA, 'CUOTA DE IVA (€)', '=SUM($I${}:$I${})-$H${}'.format(
            R0_03, R1_03, FILA_BASE)),
        (FILA_TOTAL, 'TOTAL DEL PEDIDO (€)', '=$H${}+$H${}'.format(
            FILA_BASE, FILA_CUOTA)),
    ]
    for fila, etiqueta, formula in totales:
        cel = ws.cell(row=fila, column=6, value=etiqueta)
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal='right', vertical='center')
        _merge(ws, 'F{f}:G{f}'.format(f=fila))
        # el importe se ancla en H (lo que pide la SPEC) y se combina con I
        # para que quede debajo de la columna «Total (€)» (TEC-20).
        _merge(ws, 'H{f}:I{f}'.format(f=fila))
        celf = _f(ws, 'H{}'.format(fila), formula, registro, motor.FMT_EUR)
        celf.font = Font(bold=True, size=12 if fila == FILA_TOTAL else 11)
        celf.alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=FILA_TOTAL, column=8).fill = PatternFill(
        'solid', fgColor=motor.CF_VERDE_BG)

    # ---- RD-17 · ¿supera el pedido mínimo del proveedor? -----------------
    # La cabecera iba a buscar el pedido mínimo y lo ENSEÑABA, pero nada lo
    # comparaba con el total: es la comprobación que un jefe de compras hace
    # veinte veces por semana (por debajo del mínimo, portes) y estaba a una
    # fórmula, con las dos celdas ya en la misma hoja. El coste de fallar lo
    # documenta el propio kit: «Portes de 12 € por debajo del mínimo».
    for ref in [str(r) for r in ws.merged_cells.ranges]:
        primera = int(''.join(ch for ch in ref.split(':')[0] if ch.isdigit()))
        if primera == FILA_MINIMO:
            ws.unmerge_cells(ref)
    for col in range(1, 10):
        ws.cell(row=FILA_MINIMO, column=col).value = None
    cel = _f(ws, 'A{}'.format(FILA_MINIMO),
             '=IF(OR($E$4="",$H${t}=""),"",IF($H${t}>=$E$4,'
             '"✓ El pedido supera el mínimo de este proveedor",'
             '"⚠ POR DEBAJO DEL PEDIDO MÍNIMO: pueden cobrarte portes"))'
             .format(t=FILA_BASE), registro)
    cel.font = Font(bold=True)
    cel.alignment = Alignment(horizontal='left', vertical='center')
    _merge(ws, 'A{f}:E{f}'.format(f=FILA_MINIMO))
    cel = ws.cell(row=FILA_MINIMO, column=6, value='FALTA PARA EL MÍNIMO (€)')
    cel.font = Font(bold=True)
    cel.alignment = Alignment(horizontal='right', vertical='center')
    _merge(ws, 'F{f}:G{f}'.format(f=FILA_MINIMO))
    _merge(ws, 'H{f}:I{f}'.format(f=FILA_MINIMO))
    celf = _f(ws, 'H{}'.format(FILA_MINIMO),
              '=IF(OR($E$4="",$H${t}=""),"",MAX(0,$E$4-$H${t}))'
              .format(t=FILA_BASE), registro, motor.FMT_EUR)
    celf.font = Font(bold=True)
    celf.alignment = Alignment(horizontal='right', vertical='center')
    motor._limpiar_cf(ws, set(['A{f}:E{f}'.format(f=FILA_MINIMO)]))
    motor.semaforo(ws, 'A{f}:E{f}'.format(f=FILA_MINIMO),
                   [('POR DEBAJO', 'rojo'), ('supera', 'verde')])

    # ---- desglose por tipo (DOM-18/COM-16) ------------------------------
    _cabecera(ws, FILA_DESGLOSE, ['IVA %', 'Base (€)', 'Cuota (€)',
                                  'Total (€)'])
    for i, tipo in enumerate((4, 10, 21)):
        fila = FILA_DESGLOSE + 1 + i
        ws.cell(row=fila, column=1, value=tipo).number_format = motor.FMT_ENT
        _f(ws, 'B{}'.format(fila),
           '=SUMIF($H${a}:$H${b},$A{f},$G${a}:$G${b})'.format(
               a=R0_03, b=R1_03, f=fila), registro, motor.FMT_EUR)
        _f(ws, 'C{}'.format(fila),
           '=IFERROR($B{f}*$A{f}/100,"")'.format(f=fila), registro,
           motor.FMT_EUR)
        _f(ws, 'D{}'.format(fila), '=$B{f}+$C{f}'.format(f=fila), registro,
           motor.FMT_EUR)
    _nota(ws, FILA_DESGLOSE + 4,
          'Desglose por tipo de IVA: es lo que pide una factura y lo que te '
          'permite cuadrar el pedido con el albarán. La suma de la columna '
          'Base tiene que coincidir con la BASE IMPONIBLE de arriba.', 4)

    informe.append('03:Pedido Actual: IVA por línea desde la categoría, doble '
                   'guarda en el subtotal, totales en H40:H42 y desglose por '
                   'tipo en A45:D48 (DOM-07/DOM-18/TEC-13/TEC-20/COM-16)')


def _proveedores_03(wb, informe, registro):
    """§7-bis.2 — hoja VISIBLE de 20 filas: es lo que hace útil el desplegable
    y lo que alimenta el `VLOOKUP` de la cabecera del pedido (DOM-19/COM-17)."""
    ws = _nueva_hoja(wb, 'Proveedores', despues='Pedido Actual')
    _anchos(ws, [30, 22, 16, 28, 16, 26, 12, 40])
    _titulo(ws, 1, 'PROVEEDORES', 8)
    _nota(ws, 2, 'Da de alta aquí a tus proveedores: el desplegable de la '
                 'hoja Pedido Actual y el teléfono y el pedido mínimo de su '
                 'cabecera salen de esta tabla. Las seis primeras filas son '
                 'un ejemplo: písalas con las tuyas. ESTA TABLA ES LA COPIA '
                 'OPERATIVA del Directorio de Proveedores del fichero 02: los '
                 'nueve libros del kit son ficheros independientes y no se '
                 'enlazan entre sí a propósito (un .xlsx movido de carpeta '
                 'rompería la referencia y verías #¡REF!), así que si cambias '
                 'allí un pedido mínimo o un día de pedido, cámbialo también '
                 'aquí: es el que de verdad alimenta el pedido.', 8)
    _cabecera(ws, 3, ['Proveedor', 'Categoría principal', 'Teléfono', 'Email',
                      'Pedido mínimo (€)', 'Día de pedido',
                      'Plazo de entrega (días)', 'Notas'])
    _cebra(ws, 4, 23, 8)
    for i, fila_datos in enumerate(PROVEEDORES):
        fila = 4 + i
        for j, valor in enumerate(fila_datos):
            ws.cell(row=fila, column=1 + j).value = valor
        ws.cell(row=fila, column=1).value = fila_datos[0] + ' (ejemplo)'
    _dv(ws, 'B4:B23', _lista_inline(motor.CATEGORIAS), 'Categoría no válida',
        'Las 10 categorías del kit. Son las mismas en las 9 plantillas.')
    # Esta nota NO es decorativa: `motor._rango_filas` recorta el bloque a
    # `ws.max_row` cuando la hoja no lo supera, y con la fila 23 en blanco el
    # verde y el desplegable se quedaban en la 22 — la última de las 20 filas
    # prometidas habría salido bloqueada.
    _nota(ws, 25,
          'La ficha COMPLETA de cada proveedor (CIF, nº RGSEAA, condiciones '
          'de pago, homologación y comparativa de precios) va en el fichero '
          '02 del kit. Aquí basta con lo que necesita el pedido.', 8)
    informe.append('03:Proveedores: hoja nueva VISIBLE, 20 filas y 6 de '
                   'ejemplo (§1.6, §7-bis.2, DOM-19/COM-17)')


def _historial_03(wb, informe, registro):
    """La primera fila queda enlazada al pedido en curso: registrar un pedido
    pasa de teclearlo todo otra vez a copiar una fila (DOM-18/TEC-20)."""
    ws = _nueva_hoja(wb, 'Historial Pedidos', despues='Proveedores')
    _anchos(ws, [13, 16, 26, 16, 14, 14, 14, 22, 34])
    _titulo(ws, 1, 'HISTORIAL DE PEDIDOS', 9)
    _nota(ws, 2, 'La fila 4 está ENLAZADA al pedido en curso. Cuando lo '
                 'cierres, copia esa fila y pégala como valor en la primera '
                 'fila libre; después ya puedes vaciar el pedido.', 9)
    _cabecera(ws, 3, ['Fecha', 'Nº Pedido', 'Proveedor', 'Productos',
                      'Subtotal (€)', 'IVA (€)', 'Total (€)', 'Estado',
                      'Observaciones'])
    _cebra(ws, 4, 43, 9)

    # Una referencia a una celda VACÍA devuelve 0 en Excel: sin la guarda, el
    # historial de un pedido a medio escribir enseñaba «0» en Nº Pedido y en
    # Proveedor y «00/01/1900» en Fecha. Los tres importes (E/F/G) sí van
    # directos, como pide la SPEC: un pedido vacío vale 0,00 € y eso es cierto.
    enlaces = [('A4', "=IF('Pedido Actual'!$C$4=\"\",\"\","
                      "'Pedido Actual'!$C$4)"),
               ('B4', "=IF('Pedido Actual'!$C$6=\"\",\"\","
                      "'Pedido Actual'!$C$6)"),
               ('C4', "=IF('Pedido Actual'!$C$3=\"\",\"\","
                      "'Pedido Actual'!$C$3)"),
               ('D4', '=IF(COUNTIF(\'Pedido Actual\'!$B$9:$B$38,"<>")=0,"",'
                      'COUNTIF(\'Pedido Actual\'!$B$9:$B$38,"<>")&" líneas")'),
               ('E4', "='Pedido Actual'!$H${}".format(FILA_BASE)),
               ('F4', "='Pedido Actual'!$H${}".format(FILA_CUOTA)),
               ('G4', "='Pedido Actual'!$H${}".format(FILA_TOTAL))]
    for coord, formula in enlaces:
        _f(ws, coord, formula, registro)
    ws['E4'].number_format = motor.FMT_EUR
    ws['F4'].number_format = motor.FMT_EUR
    ws['G4'].number_format = motor.FMT_EUR
    ws['A4'].number_format = motor.FMT_FECHA
    ws['I4'] = ('PEDIDO EN CURSO — esta fila se actualiza sola y NO SUMA en '
                'el total del periodo. Cópiala como valor en la primera fila '
                'libre cuando cierres el pedido.')

    for i, datos in enumerate(HISTORIAL):
        fila = 5 + i
        fecha, num, prov, prod, base, cuota, total, estado, obs = datos
        # RD-28/RT-20 · fecha CONSTANTE, no `TODAY()-n`: son celdas verdes
        # (dato del cliente) y un historial que se recoloca solo cada vez que
        # abres el libro no es un historial.
        ws.cell(row=fila, column=1,
                value=fecha).number_format = motor.FMT_FECHA
        ws.cell(row=fila, column=2).value = num
        ws.cell(row=fila, column=3).value = prov
        ws.cell(row=fila, column=4).value = prod
        ws.cell(row=fila, column=5,
                value=base).number_format = motor.FMT_EUR
        ws.cell(row=fila, column=6,
                value=cuota).number_format = motor.FMT_EUR
        ws.cell(row=fila, column=7,
                value=total).number_format = motor.FMT_EUR
        ws.cell(row=fila, column=8).value = estado
        ws.cell(row=fila, column=9).value = obs + ' (ejemplo)'

    _dv(ws, 'H4:H43', _lista_inline(ESTADOS_PEDIDO), 'Estado no válido',
        'Estado del pedido. Sirve para saber de un vistazo qué está pendiente '
        'de recibir y qué está pendiente de factura.')

    # totales del periodo, por debajo del bloque de datos.
    #
    # RT-10 · arrancan en la fila 5, NO en la 4. La 4 es la fila enlazada al
    # pedido en curso, y el flujo que documenta la propia hoja es «copia esa
    # fila y pégala como valor en la primera fila libre»: en el momento en que
    # el cliente lo hace —y hasta que además vacíe el Pedido Actual, que es un
    # segundo paso separado— el pedido recién cerrado quedaba contado DOS
    # veces en el total del periodo, sin ningún aviso.
    cel = ws.cell(row=45, column=4, value='TOTAL DEL PERIODO (filas 5 a 43)')
    cel.font = Font(bold=True)
    cel.alignment = Alignment(horizontal='right')
    for col in ('E', 'F', 'G'):
        celf = _f(ws, '{}45'.format(col),
                  '=SUM(${c}$5:${c}$43)'.format(c=col), registro,
                  motor.FMT_EUR)
        celf.font = Font(bold=True)

    # verde a mano: esta hoja está en SIN_VERDE_AUTO (ver cabecera del módulo)
    motor.marcar_verde(ws, 'H4:I4')
    motor.marcar_verde(ws, 'A5:I43')
    informe.append('03:Historial Pedidos: 40 filas, fila 4 enlazada al pedido '
                   'en curso (E4/F4/G4 = H40/H41/H42), DV de estado, 3 filas '
                   'de ejemplo y total del periodo (DOM-18/§1.4)')


def _instrucciones_03(wb, informe):
    ws = wb['Instrucciones']
    ancla = _instrucciones(ws, [
        'Pedidos de Compra',
        'AI Chef Pro — aichef.pro',
        'Genera pedidos de compra con el IVA correcto en cada línea y con el '
        'desglose que pide un albarán.',
        '1. Rellena una sola vez tus datos de emisor en el bloque verde de la '
        'derecha: establecimiento, CIF, dirección de entrega y contacto. Sin '
        'ellos, quien recibe el pedido no sabe quién pide ni dónde entregar.',
        '2. Da de alta a tus proveedores en la hoja \'Proveedores\'. El '
        'desplegable del pedido, el teléfono y el pedido mínimo de la '
        'cabecera salen de esa tabla.',
        '3. En \'Pedido Actual\' escribe producto, categoría, unidad, '
        'cantidad y precio. El IVA se rellena solo: primero busca el PRODUCTO '
        'en la tabla de excepciones de la hoja Listas y, si no está, aplica '
        'el tipo de su categoría. Puedes cambiarlo línea a línea con el '
        'desplegable 4 / 10 / 21. Si te dejas la categoría, la casilla del '
        'IVA se queda EN BLANCO en vez de caer al 10 % en silencio: en una '
        'línea de alcohol o de limpieza eso serían 11 puntos de menos.',
        '3 bis. Las TRES primeras líneas son un ejemplo con cantidad y precio '
        'para que veas el pedido funcionando —base, cuota, total y desglose '
        'por los tres tipos—. BÓRRALAS antes de enviar tu primer pedido.',
        '4. Si escribes cantidad y te dejas el precio, la casilla del '
        'subtotal avisa con «falta coste» en vez de valorar la línea en '
        '0,00 € y sumarla al total como si fuera gratis.',
        '5. Los totales del pedido están en las filas 40 a 42: base '
        'imponible, cuota de IVA y total. Justo debajo, en la fila 43, el '
        'fichero te dice si el pedido SUPERA EL PEDIDO MÍNIMO de ese '
        'proveedor y cuánto te falta si no: es lo que evita que te cobren '
        'portes. Y en las filas 45 a 48 tienes el desglose por tipo (4 %, '
        '10 % y 21 %).',
        '6. Cuando cierres el pedido, copia la fila 4 de \'Historial '
        'Pedidos\' y pégala como valor: ya viene enlazada al pedido en curso. '
        'Esa fila 4 NO suma en el total del periodo, precisamente para que el '
        'pedido no se cuente dos veces mientras haces la copia.',
        '7. La hoja \'Listas\' guarda DOS tablas: el tipo de IVA de cada '
        'categoría y, al lado, las excepciones por producto (los cereales, '
        'las harinas y las legumbres van al 4 % aunque su categoría sea del '
        '10 %; los refrescos con azúcares añadidos, al 21 %). Edítalas si tu '
        'caso es distinto: son tablas, no números escondidos dentro de la '
        'fórmula.',
        'Para enviarlo: Archivo → Imprimir → PDF. La hoja ya está en A4 '
        'apaisado y repite la cabecera de la tabla en cada página.',
    ])
    informe.append('03:Instrucciones: reescritas; retirada la promesa de la '
                   'pestaña Imprimible (§1.6) y corregido el nombre de '
                   'Historial Pedidos; ancla de versión en A{}'.format(ancla))


# ==========================================================================
# 04 · Recepción de mercancías
# ==========================================================================
#: Ronda 2 · el layout crece de 17 a 22 columnas. Lo que entra y por qué:
#:   E «Categoría (kit)»           RD-23/RC-06 · el 04 era el ÚNICO fichero
#:                                 que no hablaba la taxonomía de los otros 8
#:   F «Familia sugerida»          RD-23 · el puente categoría → familia
#:   L «Precio/ud (€)»             RD-16 · sin él la recepción calculaba la
#:   M «Valor de la diferencia»            diferencia en UNIDADES y el dinero
#:                                         aparecía dos hojas después,
#:                                         tecleado a mano y sin relación
#:   P «Tª mín. (°C)» (oculta)     RD-06/RT-15 · la conformidad sólo tenía
#:                                 techo: un pescado a -20 °C salía CONFORME
CAB_04 = ['Fecha', 'Proveedor', 'Nº albarán/factura', 'Producto',
          'Categoría (kit)', 'Familia sugerida', 'Familia', 'Nº de lote',
          'Pedido (ud)', 'Recibido (ud)', 'Diferencia (ud)', 'Precio/ud (€)',
          'Valor de la diferencia (€)', 'Caducidad', 'Temp. °C',
          'Tª mín. (°C)', 'Tª máx. (°C)', 'Conforme (Tª)', 'Calidad visual',
          'Incidencia', 'Receptor', 'Firma']
ANCHOS_04 = [12, 24, 16, 26, 18, 32, 32, 14, 11, 12, 13, 13, 17, 12, 10, 11,
             11, 20, 13, 34, 18, 14]
R0_04, R1_04 = 4, 43
FILA_RESUMEN_04 = 45

#: La familia EFECTIVA de la línea: la que el usuario haya afinado en G y, si
#: no ha afinado nada, la que propone el puente en F.
_FAM = 'IF($G{f}="",$F{f},$G{f})'


def _recepcion(wb, informe, registro):
    _temperaturas_04(wb, informe, registro)
    _control_recepcion_04(wb, informe, registro)
    _incidencias_04(wb, informe, registro)
    _instrucciones_04(wb, informe)


def _temperaturas_04(wb, informe, registro):
    """DOM-05/COM-08 — de tabla de TEXTO a tabla NUMÉRICA y legal.

    Ronda 2 (RD-06/RT-01/RT-14/RT-15/RD-23): la tabla gana una columna de
    MÍNIMO —sin ella la conformidad sólo tenía techo y un pescado recibido a
    -20 °C salía «CONFORME»—, corrige el umbral de los congelados (la nota
    citaba la tolerancia de -15 °C y el número decía -18, así que rechazaba lo
    que la norma admite), añade las dos familias que faltaban (bebidas y no
    alimentario: el 01 precarga 8 bebidas y 13 productos de limpieza que no
    tenían dónde ir) y publica el PUENTE categoría → familia.
    """
    ws = _nueva_hoja(wb, 'Verificación Temperaturas')
    _anchos(ws, [46, 44, 14, 14, 52, 4, 24, 40])
    _titulo(ws, 1, 'TEMPERATURAS DE RECEPCIÓN — LÍMITES LEGALES', 5)
    _cabecera(ws, 3, ['Familia', 'Tª ideal', 'Tª mín. aceptable (°C)',
                      'Tª máx. aceptable (°C)', 'Base normativa'])
    _cebra(ws, 4, R1_TEMP, 5)
    for i, datos in enumerate(TEMPERATURAS):
        familia, ideal, minima, maxima, norma = datos
        fila = 4 + i
        ws.cell(row=fila, column=1).value = familia
        ws.cell(row=fila, column=2).value = ideal
        for col, valor in ((3, minima), (4, maxima)):
            cel = ws.cell(row=fila, column=col, value=valor)
            cel.alignment = Alignment(horizontal='center')
        ws.cell(row=fila, column=5).value = norma
        for col in range(1, 6):
            ws.cell(row=fila, column=col).alignment = Alignment(
                vertical='top', wrap_text=True,
                horizontal='center' if col in (3, 4) else 'general')

    # ---- RD-23 · el puente categoría del kit → familia normativa ---------
    _cabecera_puente(ws, 3, 7, ['Categoría del kit', 'Familia por defecto'])
    for i, par in enumerate(PUENTE_CATEGORIA_FAMILIA):
        fila = 4 + i
        ws.cell(row=fila, column=7).value = par[0]
        cel = ws.cell(row=fila, column=8, value=par[1])
        cel.alignment = Alignment(vertical='top', wrap_text=True)
    ws.cell(row=R1_PUENTE + 2, column=7).value = (
        'De aquí sale la columna «Familia sugerida» del registro. AFÍNALA en '
        'la columna «Familia» cuando la línea lo pida: la carne picada '
        '(2 °C), los despojos (3 °C), las aves (4 °C) y la comida preparada '
        'tienen umbral propio, y los huevos tienen su propia familia aunque '
        'su categoría de compra sea «Otros».')
    ws.cell(row=R1_PUENTE + 2, column=7).alignment = Alignment(
        vertical='top', wrap_text=True)
    ws.merge_cells(start_row=R1_PUENTE + 2, start_column=7,
                   end_row=R1_PUENTE + 5, end_column=8)

    _nota(ws, R1_TEMP + 2,
          'Las columnas «Tª mín.» y «Tª máx. aceptable» son las que compara '
          'la hoja Control Recepción: por eso son números y no «7°C máx.» '
          'como en la v1.1, donde el criterio existía en papel y no se podía '
          'aplicar a ningún dato. «N/A» significa que esa familia no se '
          'controla por temperatura en la recepción. Los congelados llevan '
          '-15 °C porque ése es el límite de RECEPCIÓN (tolerancia breve de '
          'transporte); una vez almacenados, a -18 °C. Estos umbrales '
          'documentan tu registro de recepción; no sustituyen a tu plan '
          'APPCC ni a un asesor.', 5)
    informe.append('04:Verificación Temperaturas: 6 familias de texto → {} '
                   'familias con umbral MÍNIMO y MÁXIMO numéricos y base '
                   'normativa (DOM-05/COM-08, RD-06/RT-14/RT-15/RD-23); la '
                   'carne picada baja de 7 °C a 2 °C, los congelados se '
                   'reciben hasta -15 °C y el puente categoría → familia va '
                   'en G3:H{} (RD-23)'.format(len(TEMPERATURAS), R1_PUENTE))


def _cabecera_puente(ws, fila, col0, textos):
    for i, t in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i, value=t)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)


def _control_recepcion_04(wb, informe, registro):
    ws = _nueva_hoja(wb, 'Control Recepción', despues='Instrucciones')
    _anchos(ws, ANCHOS_04)
    _titulo(ws, 1, 'CONTROL DE RECEPCIÓN DE MERCANCÍAS', len(CAB_04))
    _nota(ws, 2, 'Una línea por producto recibido. Las CUATRO primeras son un '
                 'EJEMPLO y cuentan una sola historia con el Registro de '
                 'Incidencias: una entrega corta, una rechazada por '
                 'temperatura, una con mal estado y una sin control de '
                 'temperatura. Bórralas antes de empezar.', len(CAB_04))
    _cabecera(ws, 3, CAB_04)
    _cebra(ws, R0_04, R1_04, len(CAB_04))

    tabla = "'Verificación Temperaturas'!$A$4:$E${}".format(R1_TEMP)
    puente = "'Verificación Temperaturas'!$G$4:$H${}".format(R1_PUENTE)
    for fila in range(R0_04, R1_04 + 1):
        fam = _FAM.format(f=fila)
        # RD-23 · la familia normativa se PROPONE desde la categoría del kit.
        _f(ws, 'F{}'.format(fila),
           '=IF($E{f}="","",IFERROR(VLOOKUP($E{f},{p},2,FALSE),""))'
           .format(f=fila, p=puente), registro)
        # DOM-20/TEC-17 · con «Pedido» y «Recibido» una al lado de la otra, la
        # verificación era un desplegable que se marcaba a ojo.
        _f(ws, 'K{}'.format(fila),
           '=IF(OR($I{f}="",$J{f}=""),"",$J{f}-$I{f})'.format(f=fila),
           registro)
        # RD-16 · el VALOR de lo no servido. Antes la diferencia se quedaba en
        # unidades y el jefe de compras hacía la multiplicación en la
        # calculadora del móvil, que es justo el trabajo que compró para no
        # hacer.
        _f(ws, 'M{}'.format(fila),
           '=IF(OR($K{f}="",$L{f}=""),"",$K{f}*$L{f})'.format(f=fila),
           registro, motor.FMT_EUR)
        # DOM-21 · el puente que faltaba: el umbral de la familia entra en la
        # hoja del registro, que es donde están los datos. Ronda 2: ahora son
        # DOS umbrales, mínimo y máximo.
        _f(ws, 'P{}'.format(fila),
           '=IF({fam}="","",IFERROR(VLOOKUP({fam},{t},3,FALSE),""))'
           .format(fam=fam, t=tabla), registro)
        _f(ws, 'Q{}'.format(fila),
           '=IF({fam}="","",IFERROR(VLOOKUP({fam},{t},4,FALSE),""))'
           .format(fam=fam, t=tabla), registro)
        # DOM-03/COM-04 · «se marcan automáticamente en rojo» pasa a ser
        # verdad: esta columna la colorea el formato condicional del motor.
        #
        # RT-01 · el orden de las ramas NO es cosmético. La v2.0 comparaba
        # `$K<=$L` con `$L` vacía, y en Excel TODO número es menor que
        # cualquier texto: el «<=» se cumplía siempre, así que cualquier
        # temperatura de una familia que no resolviera en la tabla legal salía
        # «✓ CONFORME». Bastaba con pegar una familia (la DV no bloquea el
        # pegado), escribirla con otra grafía o editar la tabla —cosa que las
        # propias Instrucciones invitan a hacer— para que TODAS esas líneas
        # pasaran a verde. Por eso la rama de «familia sin límite» va ANTES de
        # la comparación. Y por eso la tabla legal ya se entrega protegida.
        _f(ws, 'R{}'.format(fila),
           '=IF(OR($O{f}="",{fam}=""),"",'
           'IF($Q{f}="N/A","N/A",'
           'IF($Q{f}="","⚠ FAMILIA SIN LÍMITE",'
           'IF($O{f}<$P{f},"✗ RECHAZAR (frío)",'
           'IF($O{f}<=$Q{f},"✓ CONFORME","✗ RECHAZAR (calor)")))))'
           .format(f=fila, fam=fam), registro)

    _dv(ws, 'E{}:E{}'.format(R0_04, R1_04),
        _lista_inline(motor.CATEGORIAS), 'Categoría no válida',
        'Las 10 categorías del kit, las mismas que en las otras ocho '
        'plantillas. De aquí sale la familia normativa que se te propone.')
    _dv(ws, 'G{}:G{}'.format(R0_04, R1_04),
        "='Verificación Temperaturas'!$A$4:$A${}".format(R1_TEMP),
        'Familia no válida',
        'Sólo si quieres AFINAR la familia que se te propone: la carne '
        'picada, los despojos, las aves y la comida preparada tienen umbral '
        'propio. Si la dejas en blanco se usa la sugerida.')
    _dv(ws, 'S{}:S{}'.format(R0_04, R1_04), _lista_inline(['✓', '✗', '—']),
        'Marca no válida',
        'Estado visual y organoléptico del producto: envase, olor, color y '
        'presencia de hielo o de agua libre.')

    for i, datos in enumerate(EJEMPLOS_RECEPCION):
        fila = R0_04 + i
        (fecha, prov, alb, prod, categoria, familia, lote, pedido, recibido,
         precio, cad, temp, calidad, incidencia, receptor) = datos
        for col, valor in ((1, fecha), (2, prov), (3, alb), (4, prod),
                           (5, categoria), (7, familia), (8, lote),
                           (9, pedido), (10, recibido), (12, precio),
                           (14, cad), (15, temp), (19, calidad),
                           (20, incidencia), (21, receptor)):
            if valor == '':
                continue
            ws.cell(row=fila, column=col).value = valor
        ws.cell(row=fila, column=1).number_format = motor.FMT_FECHA
        ws.cell(row=fila, column=14).number_format = motor.FMT_FECHA
        ws.cell(row=fila, column=12).number_format = motor.FMT_EUR

    # La columna de la Tª mínima es una AUXILIAR: el usuario no la rellena ni
    # la lee, y con ella a la vista la hoja pasaba de 21 a 22 columnas
    # visibles sin ganar nada. Se oculta, como el bloque P:T del 02.
    ws.column_dimensions['P'].hidden = True

    # ---- resumen (por debajo del bloque: no lo pisa el verde del motor) ---
    for etiqueta, col_val, formula, fmt in (
            ('Entregas registradas:', 4,
             '=COUNTIF($D${a}:$D${b},"<>")'.format(a=R0_04, b=R1_04), None),
            ('Líneas a rechazar:', 8,
             '=COUNTIF($R${a}:$R${b},"*RECHAZAR*")'.format(a=R0_04, b=R1_04),
             None),
            ('Diferencia total (ud):', 11,
             '=SUM($K${a}:$K${b})'.format(a=R0_04, b=R1_04), None),
            # RD-16 · lo que de verdad reclama el jefe de compras.
            ('Valor de la diferencia (€):', 14,
             '=SUM($M${a}:$M${b})'.format(a=R0_04, b=R1_04), motor.FMT_EUR)):
        cel = ws.cell(row=FILA_RESUMEN_04, column=col_val - 2, value=etiqueta)
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal='right')
        _merge(ws, '{a}{f}:{b}{f}'.format(
            a=get_column_letter(col_val - 2), f=FILA_RESUMEN_04,
            b=get_column_letter(col_val - 1)))
        celf = _f(ws, '{}{}'.format(get_column_letter(col_val),
                                    FILA_RESUMEN_04), formula, registro, fmt)
        celf.font = Font(bold=True)
    informe.append('04:Control Recepción: 0 fórmulas → 6 columnas calculadas '
                   '(F familia sugerida, K diferencia, M valor de la '
                   'diferencia, P/Q umbrales mín. y máx. por familia, R '
                   'conforme) + categoría del kit, precio/ud, lote y albarán '
                   '+ resumen con el valor reclamable; el rojo lo pone el CF '
                   'del motor sobre O4:O43 y R4:R43 (DOM-03/04/20/21, '
                   'TEC-04/17, COM-04; RD-06/RD-16/RD-23, RT-01/RT-14/RT-15)')


def _incidencias_04(wb, informe, registro):
    ws = _nueva_hoja(wb, 'Registro Incidencias', despues='Control Recepción')
    cabeceras = ['Fecha', 'Nº albarán', 'Proveedor', 'Producto',
                 'Tipo de incidencia', 'Descripción', 'Acción tomada',
                 'Responsable', 'Importe reclamado (€)', 'Abono recibido (€)',
                 'Estado']
    _anchos(ws, [12, 14, 24, 26, 26, 38, 34, 20, 18, 18, 20])
    _titulo(ws, 1, 'REGISTRO DE INCIDENCIAS — RECEPCIÓN', len(cabeceras))
    _nota(ws, 2, 'La suma de «Importe reclamado» menos la de «Abono '
                 'recibido» es lo que tu proveedor te debe. Las tres primeras '
                 'líneas son un ejemplo.', len(cabeceras))
    _cabecera(ws, 3, cabeceras, color=motor.CAB_ROJA)
    _cebra(ws, 4, 23, len(cabeceras))
    for i, datos in enumerate(EJEMPLOS_INCIDENCIA):
        fila = 4 + i
        for j, valor in enumerate(datos):
            ws.cell(row=fila, column=1 + j).value = valor
        ws.cell(row=fila, column=1).number_format = motor.FMT_FECHA
        ws.cell(row=fila, column=9).number_format = motor.FMT_EUR
        ws.cell(row=fila, column=10).number_format = motor.FMT_EUR
    _dv(ws, 'E4:E23', _lista_inline(TIPOS_INCIDENCIA),
        'Tipo de incidencia no válido',
        'Agrupar las incidencias por tipo es lo que te dice si el problema es '
        'del proveedor, del transporte o de tu propia recepción.')
    _dv(ws, 'K4:K23', _lista_inline(ESTADOS_INCIDENCIA), 'Estado no válido',
        'Mientras el estado no sea «Abono recibido» o «Cerrada», ese dinero '
        'sigue pendiente de reclamar.')

    cel = ws.cell(row=25, column=8, value='TOTALES')
    cel.font = Font(bold=True)
    cel.alignment = Alignment(horizontal='right')
    for col in ('I', 'J'):
        celf = _f(ws, '{}25'.format(col), '=SUM(${c}$4:${c}$23)'.format(c=col),
                  registro, motor.FMT_EUR)
        celf.font = Font(bold=True)
    cel = ws.cell(row=26, column=8, value='PENDIENTE DE ABONO (€)')
    cel.font = Font(bold=True)
    cel.alignment = Alignment(horizontal='right')
    celf = _f(ws, 'I26', '=IFERROR($I$25-$J$25,"")', registro, motor.FMT_EUR)
    celf.font = Font(bold=True)
    informe.append('04:Registro Incidencias: nº de albarán, importe reclamado '
                   'y abono recibido, DV de tipo y de estado, 3 ejemplos y '
                   'totales con el pendiente de abono (DOM-20/§1.4)')


def _instrucciones_04(wb, informe):
    ws = wb['Instrucciones']
    ancla = _instrucciones(ws, [
        'Recepción de Mercancías',
        'AI Chef Pro — aichef.pro',
        'Verifica cada entrega y deja registrado lo que un inspector te va a '
        'pedir: albarán, lote, temperatura, conformidad y firma.',
        '1. Cuando llegue un proveedor, abre \'Control Recepción\' y anota '
        'una línea por producto.',
        '2. Rellena el nº de albarán y el nº de lote. Son los dos datos que '
        'sostienen la trazabilidad hacia atrás: sin ellos el registro no '
        'sirve como prerrequisito documental de tu plan APPCC.',
        '3. Elige la categoría del producto —las mismas 10 del resto del '
        'kit— y la columna \'Familia sugerida\' te propone sola la familia '
        'normativa. AFÍNALA en la columna \'Familia\' cuando la línea lo '
        'pida: la carne picada (2 °C), los despojos (3 °C), las aves (4 °C) y '
        'la comida preparada tienen umbral propio, y no es lo mismo un '
        'solomillo que una hamburguesa.',
        '4. Escribe la temperatura medida: la casilla se pone ROJA y la '
        'columna de conformidad dice RECHAZAR en cuanto se sale del rango de '
        'su familia. Hay cinco respuestas posibles y conviene conocerlas: '
        'CONFORME · RECHAZAR (calor) · RECHAZAR (frío), porque un pescado a '
        '-20 °C venía congelado y una lechuga a -2 °C viene quemada por frío '
        '· N/A cuando esa familia no se controla por temperatura (huevos, '
        'secos, bebidas, no alimentario) · y FAMILIA SIN LÍMITE cuando lo que '
        'has escrito en Familia no está en la tabla legal. Esa última no es '
        'un fallo: es el fichero negándose a darte un visto bueno que no '
        'puede sostener.',
        '5. La diferencia entre lo pedido y lo recibido se calcula sola, y si '
        'pones el precio/ud también su VALOR EN EUROS: es lo que de verdad le '
        'reclamas al proveedor. Lo que no aceptes, anótalo en \'Registro '
        'Incidencias\' con ese importe; la suma de esa columna menos los '
        'abonos es lo que te deben.',
        '6. La hoja \'Verificación Temperaturas\' cita la norma de cada '
        'límite y trae, a la derecha, el puente entre las 10 categorías de '
        'compra del kit y las familias normativas. Edítala sólo si tu '
        'autoridad sanitaria te indica otra cosa; está protegida sin '
        'contraseña para que no se borre de un clic, porque de ella depende '
        'el veredicto de todas las líneas.',
        'Umbrales tomados del Reglamento (CE) 853/2004 (Anexo III), el RD '
        '3484/2000 y el Reglamento (CE) 589/2008. Documentan tus registros; '
        'no sustituyen a tu plan APPCC ni a un asesor.',
    ])
    informe.append('04:Instrucciones: la línea del rojo automático pasa a ser '
                   'CIERTA (COM-04); ancla de versión en A{}'.format(ancla))


# ==========================================================================
# 05 · Control de mermas
# ==========================================================================
R0_05, R1_05 = 4, 103
FILA_TOTAL_05 = 105
FILA_TOTAL_ANALISIS = 14


def _mermas(wb, informe, registro):
    _registro_mermas_05(wb, informe, registro)
    _analisis_05(wb, informe, registro)
    _dashboard_05(wb, informe, registro)
    _plan_accion_05(wb, informe, registro)
    _instrucciones_05(wb, informe)


def _registro_mermas_05(wb, informe, registro):
    ws = wb['Registro Diario Mermas']
    _limpiar_dv(ws)
    anadidas = motor.expandir_filas(ws, 53, R1_05, cola=(55,))
    if anadidas:
        informe.append('05:Registro Diario Mermas: {} filas nuevas (50 → 100); '
                       'el TOTAL MES baja a la fila {} y su SUM cubre el '
                       'rango completo'.format(anadidas, FILA_TOTAL_05))
    ws['F3'] = 'Precio/ud (€)'
    # el TOTAL MES vive por DEBAJO del bloque de datos, así que el §1.7 del
    # motor no le llega: se le pone el euro a mano.
    ws['G{}'.format(FILA_TOTAL_05)].number_format = motor.FMT_EUR
    _nota(ws, 2, 'Las tres primeras líneas son un EJEMPLO para que el '
                 'análisis y el dashboard arranquen con números: bórralas '
                 'antes de empezar tu mes.', 10)
    for fila in range(R0_05, R1_05 + 1):
        # §1.8 — sin la doble guarda, una merma sin precio se valoraba en
        # 0,00 € y el total del mes decía que no habías perdido nada.
        _f(ws, 'G{}'.format(fila), motor.guarda_doble('D', 'F', fila),
           registro, motor.FMT_EUR)
    _dv(ws, 'H{}:H{}'.format(R0_05, R1_05), _lista_inline(MOTIVOS),
        'Motivo no válido',
        'El motivo es lo que convierte el registro en un plan de acción: el '
        'análisis cuenta cuántas veces se repite cada uno.')
    for i, datos in enumerate(EJEMPLOS_MERMA):
        fila = R0_05 + i
        producto, categoria, cantidad, unidad, precio, motivo, quien, acc = datos
        ws.cell(row=fila, column=2).value = producto
        ws.cell(row=fila, column=3).value = categoria
        ws.cell(row=fila, column=4).value = cantidad
        ws.cell(row=fila, column=5).value = unidad
        ws.cell(row=fila, column=6, value=precio).number_format = motor.FMT_EUR
        ws.cell(row=fila, column=8).value = motivo
        ws.cell(row=fila, column=9).value = quien
        ws.cell(row=fila, column=10).value = acc
    informe.append('05:Registro Diario Mermas: doble guarda en el coste, DV '
                   'de motivo y 3 líneas de ejemplo (TEC-06/TEC-09)')


def _analisis_05(wb, informe, registro):
    """DOM-11/TEC-09/COM-12 — la hoja tenía las 7 categorías viejas y CERO
    fórmulas: la agregación era un SUMIF que nadie había escrito."""
    ws = _nueva_hoja(wb, 'Análisis por Categoría',
                     despues='Registro Diario Mermas')
    _anchos(ws, [24, 16, 18, 14, 22, 24, 4, 30, 14])
    _titulo(ws, 1, 'ANÁLISIS DE MERMAS POR CATEGORÍA', 6)
    _cabecera(ws, 3, ['Categoría', 'Nº Incidencias', 'Coste Total (€)',
                      '% del Total', 'Coste mes anterior (€)',
                      'Tendencia vs mes anterior (%)'])
    _cebra(ws, 4, 13, 6)
    reg = "'Registro Diario Mermas'"
    for i, categoria in enumerate(motor.CATEGORIAS):
        fila = 4 + i
        ws.cell(row=fila, column=1).value = categoria
        _f(ws, 'B{}'.format(fila),
           '=COUNTIF({r}!$C${a}:$C${b},$A{f})'.format(r=reg, a=R0_05, b=R1_05,
                                                      f=fila), registro,
           motor.FMT_ENT)
        _f(ws, 'C{}'.format(fila),
           '=SUMIF({r}!$C${a}:$C${b},$A{f},{r}!$G${a}:$G${b})'.format(
               r=reg, a=R0_05, b=R1_05, f=fila), registro, motor.FMT_EUR)
        _f(ws, 'D{}'.format(fila),
           '=IFERROR($C{f}/$C${t},"")'.format(f=fila, t=FILA_TOTAL_ANALISIS),
           registro, motor.FMT_PCT1)
        _f(ws, 'F{}'.format(fila),
           '=IFERROR(IF($E{f}="","",$C{f}/$E{f}-1),"")'.format(f=fila),
           registro, motor.FMT_PCT1)

    cel = ws.cell(row=FILA_TOTAL_ANALISIS, column=1, value='TOTAL DEL MES')
    cel.font = Font(bold=True)
    for col, formula, fmt in (
            ('B', '=SUM($B$4:$B$13)', motor.FMT_ENT),
            ('C', '=SUM($C$4:$C$13)', motor.FMT_EUR),
            ('D', '=IFERROR(SUM($D$4:$D$13),"")', motor.FMT_PCT1),
            ('E', '=SUM($E$4:$E$13)', motor.FMT_EUR),
            ('F', '=IFERROR(IF($E$14=0,"",$C$14/$E$14-1),"")',
             motor.FMT_PCT1)):
        celf = _f(ws, '{}{}'.format(col, FILA_TOTAL_ANALISIS), formula,
                  registro, fmt)
        celf.font = Font(bold=True)

    # ---- bloque auxiliar de motivos (pycel no implementa MODE) -----------
    # El bloque llega hasta la fila 13 (no hasta la 10): el verde del motor
    # alcanza toda la altura del bloque de datos, así que si el `COUNTIF` se
    # quedara en la 10, las tres casillas verdes de debajo invitarían a añadir
    # un motivo que nadie contaría.
    _cabecera_aux(ws, 3, 8, ['Motivo', 'Nº de veces'])
    for i, motivo in enumerate(MOTIVOS):
        fila = 4 + i
        ws.cell(row=fila, column=8).value = motivo
        _f(ws, 'I{}'.format(fila),
           '=COUNTIF({r}!$H${a}:$H${b},$H{f})'.format(r=reg, a=R0_05, b=R1_05,
                                                      f=fila), registro,
           motor.FMT_ENT)
    informe.append('05:Análisis por Categoría: 7 categorías sueltas → las 10 '
                   'canónicas con COUNTIF/SUMIF, % del total, tendencia vs '
                   'mes anterior y bloque de motivos en H4:I10 '
                   '(DOM-11/TEC-09/COM-12)')


def _cabecera_aux(ws, fila, col0, textos):
    for i, t in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i, value=t)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)


def _dashboard_05(wb, informe, registro):
    """DOM-23/COM-12 — el objetivo estrella del producto («mermas por debajo
    del 3 % sobre compras») era INCALCULABLE: no existía ninguna celda donde
    escribir las compras del mes. Y el objetivo alternativo, «< 500 €», sólo
    vale para un tamaño de negocio."""
    ws = _nueva_hoja(wb, 'Dashboard Mermas', despues='Análisis por Categoría')
    _anchos(ws, [34, 22, 22, 26])
    _titulo(ws, 1, 'DASHBOARD DE MERMAS — MES', 4)
    # «Valor» NO puede ser la cabecera: `motor.aplicar_formatos` la reconoce
    # como columna de importe y le pone `#,##0.00 €` a toda la columna. Aquí
    # conviven un euro (B4), un porcentaje (B5), un recuento (B6) y dos textos
    # (B7/B8) — medido: el 1,5 % de mermas salía impreso como «0,01 €».
    _cabecera(ws, 3, ['KPI', 'Resultado', 'Objetivo', 'Estado'])
    ana = "'Análisis por Categoría'"
    reg = "'Registro Diario Mermas'"

    ws['A4'] = 'Coste total mermas (€)'
    ws['B4'].number_format = motor.FMT_EUR
    _f(ws, 'B4', "={r}!$G${t}".format(r=reg, t=FILA_TOTAL_05), registro,
       motor.FMT_EUR)
    _f(ws, 'C4', '=IFERROR(IF($B$10="","",$B$10*$B$11),"")', registro,
       motor.FMT_EUR)
    _f(ws, 'D4', '=IF(OR($B$4="",$C$4=""),"",IF($B$4<=$C$4,"🟢 OK",'
                 '"🔴 ALERTA"))', registro)

    ws['A5'] = '% mermas sobre compras'
    _f(ws, 'B5', '=IFERROR($B$4/$B$10,"")', registro, motor.FMT_PCT1)
    _f(ws, 'C5', '=IF($B$11="","",$B$11)', registro, motor.FMT_PCT1)
    _f(ws, 'D5', '=IF(OR($B$5="",$C$5=""),"",IF($B$5<=$C$5,"🟢 OK",'
                 'IF($B$5<=$C$5*1.5,"🟡 REVISAR","🔴 ALERTA")))', registro)

    ws['A6'] = 'Nº de incidencias'
    _f(ws, 'B6', '=COUNTIF({r}!$B${a}:$B${b},"<>")'.format(r=reg, a=R0_05,
                                                           b=R1_05), registro,
       motor.FMT_ENT)
    # RT-19 · era un literal en una celda BLOQUEADA dentro de una hoja
    # protegida: el cliente no podía cambiar su propio objetivo sin
    # desproteger. En la misma hoja, el objetivo de % de mermas (B11) sí era
    # verde editable. Se aplica el mismo criterio.
    _verde(ws['C6'], motor.FMT_ENT)
    ws['C6'] = 20
    _f(ws, 'D6', '=IF($B$6="","",IF($B$6<=$C$6,"🟢 OK","🟡 REVISAR"))',
       registro)

    # RD-09/RT-04 · sin guarda de CERO, INDEX/MATCH(MAX(...)) sobre una
    # columna de SUMIF/COUNTIF que valen 0 encuentra el 0 y devuelve SIEMPRE
    # la primera fila. En cuanto el cliente hace lo que la propia hoja le
    # manda —«borra las tres líneas de ejemplo antes de empezar tu mes»— el
    # dashboard se ponía a afirmar que su categoría con más merma era
    # «Cárnicos» y su motivo más frecuente «Caducidad superada» con 0,00 € de
    # merma y 0 incidencias. O desconfías del kit entero, o te vas a por los
    # cárnicos sin motivo.
    ws['A7'] = 'Categoría con más merma'
    _f(ws, 'B7', '=IF(MAX({a}!$C$4:$C$13)<=0,"— sin mermas registradas —",'
                 'IFERROR(INDEX({a}!$A$4:$A$13,MATCH(MAX({a}!$C$4:$C$13),'
                 '{a}!$C$4:$C$13,0)),""))'.format(a=ana), registro)
    ws['C7'] = 'Atácala en el plan'

    # RT-22 · y el recuento al lado: con pocas mermas al mes media docena de
    # motivos empatan a 1 o 2 y el ranking se queda con el primero de la lista
    # auxiliar sin decir nada. Con el número delante, el empate se ve.
    ws['A8'] = 'Motivo más frecuente'
    _f(ws, 'B8', '=IF(MAX({a}!$I$4:$I$13)<=0,"— sin mermas registradas —",'
                 'IFERROR(INDEX({a}!$H$4:$H$13,MATCH(MAX({a}!$I$4:$I$13),'
                 '{a}!$I$4:$I$13,0))&" · "&MAX({a}!$I$4:$I$13)&" veces",""))'
       .format(a=ana), registro)
    ws['C8'] = 'Atácalo en el plan'
    ws['D8'] = ('Si varios motivos empatan, enseña el primero de la lista: '
                'mira la tabla de motivos del análisis para verlos todos.')

    # ---- entradas del cliente (fuera del bloque de KPI) -----------------
    ws['A10'] = 'Compras del mes (€, sin IVA)'
    ws['A10'].font = Font(bold=True)
    _verde(ws['B10'], motor.FMT_EUR)
    # RD-12 · la ÚNICA casilla de entrada del dashboard se entregaba vacía, y
    # de ella colgaban el porcentaje de mermas, el objetivo en euros y los dos
    # semáforos: al abrir el fichero se veían cuatro celdas en blanco donde el
    # cliente espera el KPI que le vendieron. El 07 hace lo contrario en el
    # mismo kit (siembra sus seis entradas y abre funcionando). Los 10.400 €
    # son las compras de enero del 07: con 79,10 € de merma sale un 0,76 %,
    # verde contra el objetivo del 3 %.
    ws['B10'] = 10400.00
    ws['A11'] = 'Objetivo de mermas sobre compras'
    ws['A11'].font = Font(bold=True)
    _verde(ws['B11'], motor.FMT_PCT1)
    ws['B11'] = 0.03
    _nota(ws, 13,
          'Escribe en la casilla verde las compras del mes SIN IVA (la base '
          'imponible de tus facturas de proveedor). Sin ese dato, el objetivo '
          'del 3 % que el propio kit fija no se puede calcular: por eso el '
          'objetivo en euros ya no es un «menos de 500 €» fijo, sino ese '
          'porcentaje sobre TUS compras.', 4)
    informe.append('05:Dashboard Mermas: 5 KPI con fórmula, entradas verdes '
                   'B10 (compras del mes) y B11 (objetivo, 3 %) y objetivo '
                   'relativo en C4; el semáforo de D lo colorea el motor '
                   '(DOM-23/COM-12)')


def _plan_accion_05(wb, informe, registro):
    """COM-10 — la hoja contenía únicamente su título: ni cabeceras, ni filas,
    ni fórmulas. Se vendía como entregable en la tarjeta del grid."""
    ws = _nueva_hoja(wb, 'Plan de Acción', despues='Dashboard Mermas')
    cabeceras = ['Prioridad', 'Categoría', 'Problema detectado', 'Causa raíz',
                 'Acción correctora', 'Responsable', 'Fecha límite',
                 'Coste mensual evitado (€)', 'Estado y seguimiento']
    _anchos(ws, [11, 20, 40, 44, 52, 24, 14, 20, 24])
    _titulo(ws, 1, 'PLAN DE ACCIÓN — REDUCCIÓN DE MERMAS', len(cabeceras))
    _nota(ws, 2, 'Las cinco primeras filas son las causas de merma que más se '
                 'repiten en cocina, ya escritas. Adáptalas, ponles fecha y '
                 'responsable, y estima tú el coste que evitas al mes.',
          len(cabeceras))
    _cabecera(ws, 3, cabeceras)
    _cebra(ws, 4, 23, len(cabeceras))
    for i, datos in enumerate(PLAN_ACCION):
        fila = 4 + i
        prioridad, categoria, problema, causa, accion, quien, ahorro, estado = datos
        for col, valor in ((1, prioridad), (2, categoria), (3, problema),
                           (4, causa), (5, accion), (6, quien), (8, ahorro),
                           (9, estado)):
            cel = ws.cell(row=fila, column=col, value=valor)
            cel.alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=fila, column=8).number_format = motor.FMT_EUR
    _dv(ws, 'A4:A23', _lista_inline(PRIORIDADES), 'Prioridad no válida',
        'Prioriza por dinero, no por comodidad: mira antes la columna «Coste '
        'Total» del análisis por categoría.')
    _dv(ws, 'I4:I23', _lista_inline(ESTADOS_PLAN), 'Estado no válido',
        'Un plan sin seguimiento es una lista de buenas intenciones.')
    cel = ws.cell(row=25, column=7, value='TOTAL AHORRO ESTIMADO')
    cel.font = Font(bold=True)
    cel.alignment = Alignment(horizontal='right')
    celf = _f(ws, 'H25', '=SUM($H$4:$H$23)', registro, motor.FMT_EUR)
    celf.font = Font(bold=True)
    informe.append('05:Plan de Acción: de título suelto a 9 columnas con DV '
                   'de prioridad y estado, 5 causas precargadas y total del '
                   'ahorro estimado (COM-10/§1.4)')


def _instrucciones_05(wb, informe):
    ws = wb['Instrucciones']
    ancla = _instrucciones(ws, [
        'Control de Mermas',
        'AI Chef Pro — aichef.pro',
        'Registra cada merma, mira dónde pierdes el dinero y ponle un plan '
        'con responsable y fecha.',
        '1. Anota cada merma en \'Registro Diario Mermas\': fecha, producto, '
        'categoría, cantidad, unidad, precio y motivo.',
        '2. El coste de la línea se calcula solo. Si te falta el precio, la '
        'celda avisa en vez de valorar la merma en 0,00 € y hacerte creer que '
        'no has perdido nada.',
        '3. \'Análisis por Categoría\' agrega por las 10 categorías del kit y '
        'cuenta los motivos. No teclees nada ahí salvo el coste del mes '
        'anterior, que es lo que da la tendencia.',
        '4. En \'Dashboard Mermas\' escribe las compras del mes SIN IVA en la '
        'casilla verde. Sin ese dato el objetivo del 3 % no se puede '
        'calcular; el porcentaje objetivo también es editable.',
        '5. \'Plan de Acción\' convierte el análisis en tareas. Trae cinco '
        'causas típicas ya escritas para que las adaptes a tu casa.',
        'OBJETIVO: mantener las mermas por debajo del 3 % sobre las compras '
        'del mes. Es un porcentaje sobre TUS compras, no una cifra fija en '
        'euros: 500 € de merma son mucho en un bar y poco en un hotel.',
    ])
    informe.append('05:Instrucciones: nombres de hoja corregidos (Registro '
                   'Diario Mermas, Dashboard Mermas) y objetivo explicado; '
                   'ancla de versión en A{}'.format(ancla))


# ==========================================================================
# Demostraciones con pycel (SPEC §5)
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            valor = xl.evaluate(ref)
        except Exception as e:                                  # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)
    if hasattr(valor, 'item'):
        try:
            return valor.item()
        except Exception:                                       # noqa: BLE001
            return valor
    return valor


def _copia(carpeta, fname, dir_demos, sufijo):
    """Copia DESECHABLE y desprotegida: los casos de prueba escriben en celdas
    que en el entregable están bloqueadas."""
    dst = os.path.join(dir_demos, '{}-{}.xlsx'.format(fname[:-5], sufijo))
    shutil.copy2(os.path.join(carpeta, fname), dst)
    wb = openpyxl.load_workbook(dst)
    for ws in wb.worksheets:
        ws.protection.sheet = False
    return dst, wb


def demos(carpeta, origen):
    """Casos con pycel sobre los tres ficheros del grupo.

    Cada caso ESCRIBE una entrada en una copia desechable y comprueba la
    DIRECCIÓN del resultado. No basta con que la fórmula esté escrita: lo que
    la ronda 1 encontró es precisamente lógica que parecía existir y no
    existía.
    """
    dir_demos = os.path.join(os.path.dirname(os.path.abspath(carpeta)),
                             'demos-grupo-b')
    if os.path.isdir(dir_demos):
        shutil.rmtree(dir_demos)
    os.makedirs(dir_demos)
    fuera = {}
    fallos = []

    fuera['iva_por_categoria_03'] = _demo_iva(carpeta, dir_demos, fallos)
    fuera['doble_guarda_03'] = _demo_guarda(carpeta, dir_demos, fallos)
    fuera['conforme_por_familia_04'] = _demo_conforme(carpeta, dir_demos,
                                                      fallos)
    fuera['agregacion_mermas_05'] = _demo_mermas(carpeta, dir_demos, fallos)
    fuera['grupo_b_fallos'] = fallos
    return fuera


def _demo_iva(carpeta, dir_demos, fallos):
    """DOM-07/TEC-13/TEC-20/COM-16 — el tipo tiene que MOVERSE con la
    categoría, y la cuota y el desglose por tipo tienen que seguirlo."""
    dst, wb = _copia(carpeta, F03, dir_demos, 'iva')
    ws = wb['Pedido Actual']
    # tres líneas reales de un pedido de restaurante, una por tipo
    datos = [(9, 'Solomillo de ternera', 'Cárnicos', 'kg', 2, 32.00),
             (10, 'Vino tinto D.O.', 'Bebidas Alcohólicas', 'ud', 6, 5.00),
             (11, 'Lechuga', 'Verduras/Frutas', 'ud', 10, 1.10)]
    for fila, prod, cat, ud, cant, precio in datos:
        ws.cell(row=fila, column=2).value = prod
        ws.cell(row=fila, column=3).value = cat
        ws.cell(row=fila, column=4).value = ud
        ws.cell(row=fila, column=5).value = cant
        ws.cell(row=fila, column=6).value = precio
    wb.save(dst)
    xl = _pycel(dst)
    lineas = []
    for fila, prod, cat, _u, cant, precio in datos:
        lineas.append({
            'ref': '{}:Pedido Actual:H{}/I{}'.format(F03, fila, fila),
            'producto': prod, 'categoria': cat,
            'cantidad': cant, 'precio': precio,
            'iva_pct': _ev(xl, "'Pedido Actual'!H{}".format(fila)),
            'subtotal': _ev(xl, "'Pedido Actual'!G{}".format(fila)),
            'total': _ev(xl, "'Pedido Actual'!I{}".format(fila))})
    esperado = {'Cárnicos': 10, 'Bebidas Alcohólicas': 21,
                'Verduras/Frutas': 4}
    for linea in lineas:
        if linea['iva_pct'] != esperado[linea['categoria']]:
            fallos.append('03:Pedido Actual:{}: «{}» debería tributar al {} % '
                          'y da {}'.format(linea['ref'], linea['categoria'],
                                           esperado[linea['categoria']],
                                           linea['iva_pct']))
    base = _ev(xl, "'Pedido Actual'!H{}".format(FILA_BASE))
    cuota = _ev(xl, "'Pedido Actual'!H{}".format(FILA_CUOTA))
    total = _ev(xl, "'Pedido Actual'!H{}".format(FILA_TOTAL))
    # 64,00 + 30,00 + 11,00 = 105,00 · cuota 6,40 + 6,30 + 0,44 = 13,14
    if base is None or abs(base - 105.0) > 0.01:
        fallos.append('03:Pedido Actual:H{}: la base debería ser 105,00 € y '
                      'da {}'.format(FILA_BASE, base))
    if cuota is None or abs(cuota - 13.14) > 0.01:
        fallos.append('03:Pedido Actual:H{}: la cuota debería ser 13,14 € y '
                      'da {}'.format(FILA_CUOTA, cuota))
    desglose = []
    for i, tipo in enumerate((4, 10, 21)):
        fila = FILA_DESGLOSE + 1 + i
        desglose.append({'ref': '{}:Pedido Actual:B{}'.format(F03, fila),
                         'tipo': tipo,
                         'base': _ev(xl, "'Pedido Actual'!B{}".format(fila)),
                         'cuota': _ev(xl, "'Pedido Actual'!C{}".format(fila)),
                         'total': _ev(xl, "'Pedido Actual'!D{}".format(fila))})
    suma = sum(d['base'] for d in desglose
               if isinstance(d['base'], (int, float)))
    if abs(suma - 105.0) > 0.01:
        fallos.append('03:Pedido Actual:B46:B48: el desglose por tipo suma {} '
                      'y la base es 105,00 €'.format(suma))
    # el mismo pedido con el 10 % a mano (v1.1) habría cobrado de menos
    v11 = round(105.0 * 1.10, 2)
    return {'copia_desechable': dst,
            'ref_tabla_iva': '{}:Listas:A2:B11'.format(F03),
            'lineas': lineas,
            'base_imponible': base, 'cuota_iva': cuota, 'total': total,
            'desglose_por_tipo': desglose,
            'total_v1_1_con_el_10_a_mano': v11,
            'diferencia_eur': round((total or 0) - v11, 2),
            'lectura': 'el mismo pedido con el 10 % escrito a mano en las tres '
                       'líneas se enviaba por {} € en vez de {} €'.format(
                           v11, total)}


def _demo_guarda(carpeta, dir_demos, fallos):
    """§1.8/TEC-06 — una línea con cantidad y sin precio valía 0,00 € y se
    sumaba al total como si el proveedor la regalara."""
    dst, wb = _copia(carpeta, F03, dir_demos, 'guarda')
    ws = wb['Pedido Actual']
    ws['B12'] = 'Aceite de oliva virgen extra'
    ws['C12'] = 'Secos/Granos'
    ws['D12'] = 'L'
    ws['E12'] = 15                      # cantidad SIN precio
    wb.save(dst)
    xl = _pycel(dst)
    sin_precio = _ev(xl, "'Pedido Actual'!G12")
    wb2 = openpyxl.load_workbook(dst)
    wb2['Pedido Actual']['F12'] = 6.40
    wb2.save(dst)
    xl2 = _pycel(dst)
    con_precio = _ev(xl2, "'Pedido Actual'!G12")
    ok = (sin_precio == '⚠ falta coste' and con_precio == 96.0)
    if not ok:
        fallos.append('03:Pedido Actual:G12: la doble guarda no funciona '
                      '(sin precio → {!r}, con precio → {!r})'
                      .format(sin_precio, con_precio))
    return {'copia_desechable': dst,
            'ref': '{}:Pedido Actual:G12'.format(F03),
            'formula': motor.guarda_doble('E', 'F', 12),
            'cantidad_15_sin_precio': sin_precio,
            'cantidad_15_precio_6_40': con_precio,
            'ok': ok,
            'lectura': 'con la guarda de la v1.1 (=IF(D12="","",D12*E12)) esa '
                       'misma línea valía 0,00 € y el total del pedido mentía'}


def _demo_conforme(carpeta, dir_demos, fallos):
    """DOM-05/DOM-21/TEC-04/COM-04/COM-08 — el umbral tiene que salir de la
    FAMILIA. La v1.1 daba 7 °C a toda la carne: una picada a 6 °C se aceptaba.

    Ronda 2 · se añaden los tres casos que refutaron la v2.0: la familia que
    NO resuelve (RT-01, daba «CONFORME» a cualquier temperatura), la entrega
    demasiado FRÍA (RT-15, un pescado a -20 °C salía conforme) y el congelado
    dentro de la tolerancia de transporte (RT-14, se rechazaba a -16 °C).
    """
    dst, wb = _copia(carpeta, F04, dir_demos, 'conforme')
    ws = wb['Control Recepción']
    # (fila, familia, temperatura, veredicto esperado, umbral máx. esperado)
    casos = [
        (10, 'Carne picada', 1, '✓ CONFORME', 2),
        (11, 'Carne picada', 6, '✗ RECHAZAR (calor)', 2),
        (12, 'Canal y despiece de ungulados domésticos', 6, '✓ CONFORME', 7),
        (13, 'Congelados', -20, '✓ CONFORME', -15),
        (14, 'Congelados', -16, '✓ CONFORME', -15),        # RT-14
        (15, 'Huevos', 18, 'N/A', 'N/A'),
        (16, 'Pescado fresco y marisco', 5, '✗ RECHAZAR (calor)', 2),
        (17, 'Pescado fresco y marisco', -20, '✗ RECHAZAR (frío)', 2),  # RT-15
        (18, 'Charcutería', 18, '⚠ FAMILIA SIN LÍMITE', ''),            # RT-01
        (19, 'No alimentario (limpieza, menaje y desechables)', 25,
         'N/A', 'N/A'),                                                 # RD-23
    ]
    for fila, familia, temp, _esp, _umbral in casos:
        ws.cell(row=fila, column=4).value = 'Caso de prueba'
        ws.cell(row=fila, column=7).value = familia      # G · Familia
        ws.cell(row=fila, column=15).value = temp        # O · Temp. °C
    ws['I22'] = 12          # pedido
    ws['J22'] = 9           # recibido
    ws['L22'] = 14.00       # precio/ud
    # RD-23 · una línea SIN familia, sólo con la categoría del kit: el puente
    # tiene que proponerla y el veredicto tiene que salir igual.
    ws['D23'] = 'Caso de prueba (sólo categoría)'
    ws['E23'] = 'Pescados'
    ws['O23'] = 5
    wb.save(dst)
    xl = _pycel(dst)
    pruebas = []
    for fila, familia, temp, esperado, umbral in casos:
        obtenido = _ev(xl, "'Control Recepción'!R{}".format(fila))
        leido = _ev(xl, "'Control Recepción'!Q{}".format(fila))
        pruebas.append({'ref': '{}:Control Recepción:R{}'.format(F04, fila),
                        'familia': familia, 'temperatura_c': temp,
                        'umbral_maximo_leido': leido, 'esperado': esperado,
                        'obtenido': obtenido, 'ok': obtenido == esperado})
        if obtenido != esperado:
            fallos.append('04:Control Recepción:R{}: {} a {} °C debería dar '
                          '«{}» y da «{}»'.format(fila, familia, temp,
                                                  esperado, obtenido))
        if leido != umbral:
            fallos.append('04:Control Recepción:Q{}: el umbral de «{}» '
                          'debería ser {} y el VLOOKUP trae {}'
                          .format(fila, familia, umbral, leido))
    dif = _ev(xl, "'Control Recepción'!K22")
    if dif != -3:
        fallos.append('04:Control Recepción:K22: 12 pedidas y 9 recibidas '
                      'deberían dar -3 y da {}'.format(dif))
    valor = _ev(xl, "'Control Recepción'!M22")
    if valor != -42:
        fallos.append('04:Control Recepción:M22: 3 unidades de menos a 14,00 '
                      '€ deberían dar -42,00 € y dan {}'.format(valor))
    sugerida = _ev(xl, "'Control Recepción'!F23")
    puente_ok = sugerida == 'Pescado fresco y marisco'
    veredicto_puente = _ev(xl, "'Control Recepción'!R23")
    if not puente_ok:
        fallos.append('04:Control Recepción:F23: la categoría «Pescados» '
                      'debería proponer «Pescado fresco y marisco» y propone '
                      '«{}»'.format(sugerida))
    if veredicto_puente != '✗ RECHAZAR (calor)':
        fallos.append('04:Control Recepción:R23: con la familia SUGERIDA (sin '
                      'afinar), 5 °C de pescado debería dar «✗ RECHAZAR '
                      '(calor)» y da «{}»'.format(veredicto_puente))
    rechazos = _ev(xl, "'Control Recepción'!H{}".format(FILA_RESUMEN_04))
    return {'copia_desechable': dst,
            'ref_tabla': '{}:Verificación Temperaturas:A4:E{}'.format(
                F04, R1_TEMP),
            'pruebas': pruebas,
            'diferencia_12_pedidas_9_recibidas': dif,
            'valor_de_esa_diferencia_eur': valor,
            'familia_sugerida_desde_categoria': sugerida,
            'veredicto_con_familia_sugerida': veredicto_puente,
            'lineas_a_rechazar_en_el_resumen': rechazos,
            'lectura': 'con la tabla de la v1.1 («Carne fresca, 7°C máx.» como '
                       'TEXTO) ninguna de estas líneas se podía calcular; con '
                       'la de la v2.0, las tres últimas familias salían mal: '
                       'una familia sin límite daba CONFORME a 18 °C, un '
                       'pescado a -20 °C también, y un congelado a -16 °C '
                       '—dentro de la tolerancia que el propio fichero cita— '
                       'se rechazaba'}


def _demo_mermas(carpeta, dir_demos, fallos):
    """DOM-11/DOM-23/TEC-09/COM-12 — la agregación y el % sobre compras."""
    dst, wb = _copia(carpeta, F05, dir_demos, 'agregacion')
    ws = wb['Registro Diario Mermas']
    # dos líneas MÁS de cárnicos, sobre los 3 ejemplos que ya trae la hoja
    nuevas = [(7, 'Costillar de cerdo', 'Cárnicos', 3, 'kg', 8.50,
               'Caducidad superada'),
              (8, 'Pollo entero', 'Cárnicos', 2, 'kg', 6.50,
               'Caducidad superada')]
    for fila, prod, cat, cant, ud, precio, motivo in nuevas:
        ws.cell(row=fila, column=2).value = prod
        ws.cell(row=fila, column=3).value = cat
        ws.cell(row=fila, column=4).value = cant
        ws.cell(row=fila, column=5).value = ud
        ws.cell(row=fila, column=6).value = precio
        ws.cell(row=fila, column=8).value = motivo
    wb['Dashboard Mermas']['B10'] = 8000.0
    wb.save(dst)
    xl = _pycel(dst)

    filas = []
    for i, categoria in enumerate(motor.CATEGORIAS):
        fila = 4 + i
        filas.append({
            'ref': '{}:Análisis por Categoría:C{}'.format(F05, fila),
            'categoria': categoria,
            'incidencias': _ev(xl, "'Análisis por Categoría'!B{}".format(fila)),
            'coste_eur': _ev(xl, "'Análisis por Categoría'!C{}".format(fila)),
            'pct_del_total': _ev(xl,
                                 "'Análisis por Categoría'!D{}".format(fila))})
    # 38,40 (solomillo) + 25,50 (costillar) + 13,00 (pollo) = 76,90
    carnicos = filas[0]
    if carnicos['coste_eur'] is None or abs(carnicos['coste_eur'] - 76.90) > 0.01:
        fallos.append('05:Análisis por Categoría:C4: Cárnicos debería sumar '
                      '76,90 € y da {}'.format(carnicos['coste_eur']))
    if carnicos['incidencias'] != 3:
        fallos.append('05:Análisis por Categoría:B4: Cárnicos debería contar '
                      '3 incidencias y da {}'.format(carnicos['incidencias']))

    total = _ev(xl, "'Análisis por Categoría'!C{}".format(FILA_TOTAL_ANALISIS))
    dash = {
        'ref': '{}:Dashboard Mermas:B4:B8'.format(F05),
        'compras_del_mes_escritas_en_B10': 8000.0,
        'coste_total_mermas': _ev(xl, "'Dashboard Mermas'!B4"),
        'objetivo_eur_C4': _ev(xl, "'Dashboard Mermas'!C4"),
        'pct_sobre_compras_B5': _ev(xl, "'Dashboard Mermas'!B5"),
        'estado_D5': _ev(xl, "'Dashboard Mermas'!D5"),
        'n_incidencias_B6': _ev(xl, "'Dashboard Mermas'!B6"),
        'categoria_con_mas_merma_B7': _ev(xl, "'Dashboard Mermas'!B7"),
        'motivo_mas_frecuente_B8': _ev(xl, "'Dashboard Mermas'!B8"),
    }
    vacio = None
    if dash['objetivo_eur_C4'] is None or abs(
            dash['objetivo_eur_C4'] - 240.0) > 0.01:
        fallos.append('05:Dashboard Mermas:C4: con 8.000 € de compras y el 3 % '
                      'de objetivo, la meta son 240,00 € y da {}'
                      .format(dash['objetivo_eur_C4']))
    if dash['categoria_con_mas_merma_B7'] != 'Cárnicos':
        fallos.append('05:Dashboard Mermas:B7: la categoría con más merma '
                      'debería ser Cárnicos y da {!r}'
                      .format(dash['categoria_con_mas_merma_B7']))
    # RT-22 · el KPI ya no es sólo el nombre: lleva pegado el recuento, para
    # que el empate (que con pocas mermas al mes es la norma) se vea.
    if dash['motivo_mas_frecuente_B8'] != 'Caducidad superada · 3 veces':
        fallos.append('05:Dashboard Mermas:B8: el motivo más frecuente '
                      'debería ser «Caducidad superada · 3 veces» y da {!r}'
                      .format(dash['motivo_mas_frecuente_B8']))
    # RD-09/RT-04 · y con el registro VACÍO —que es lo que la propia hoja le
    # manda hacer al cliente— los dos KPI cualitativos tienen que callarse en
    # vez de acusar a los cárnicos con 0,00 € de merma.
    wb0 = openpyxl.load_workbook(dst)
    ws0 = wb0['Registro Diario Mermas']
    for fila in range(R0_05, 9):
        for col in range(1, 11):
            ws0.cell(row=fila, column=col).value = None
    vacio_dst = dst.replace('.xlsx', '-sin-datos.xlsx')
    wb0.save(vacio_dst)
    xl0 = _pycel(vacio_dst)
    vacio = {
        'ref': '{}:Dashboard Mermas:B7:B8 con el registro VACÍO'.format(F05),
        'copia_desechable': vacio_dst,
        'coste_total_B4': _ev(xl0, "'Dashboard Mermas'!B4"),
        'categoria_con_mas_merma_B7': _ev(xl0, "'Dashboard Mermas'!B7"),
        'motivo_mas_frecuente_B8': _ev(xl0, "'Dashboard Mermas'!B8"),
        'lectura': ('en la v2.0 estas dos celdas decían «Cárnicos» y '
                    '«Caducidad superada» con CERO mermas registradas: '
                    'MAX() de una columna de ceros es 0 y MATCH(0;…;0) '
                    'engancha la primera fila'),
    }
    esperado_vacio = '— sin mermas registradas —'
    for clave in ('categoria_con_mas_merma_B7', 'motivo_mas_frecuente_B8'):
        if vacio[clave] != esperado_vacio:
            fallos.append('05:Dashboard Mermas: con el registro vacío, {} '
                          'debería decir «{}» y dice {!r}'
                          .format(clave, esperado_vacio, vacio[clave]))
    if dash['estado_D5'] != '🟢 OK':
        fallos.append('05:Dashboard Mermas:D5: 118,50 € sobre 8.000 € es 1,5 % '
                      'y el estado debería ser OK; da {!r}'
                      .format(dash['estado_D5']))

    # segunda pasada: mismas mermas, la CUARTA parte de compras → el % se
    # dispara y el semáforo tiene que cambiar de color
    wb3 = openpyxl.load_workbook(dst)
    wb3['Dashboard Mermas']['B10'] = 2000.0
    wb3.save(dst)
    xl3 = _pycel(dst)
    dash_pobre = {
        'compras_del_mes_escritas_en_B10': 2000.0,
        'pct_sobre_compras_B5': _ev(xl3, "'Dashboard Mermas'!B5"),
        'objetivo_eur_C4': _ev(xl3, "'Dashboard Mermas'!C4"),
        'estado_D5': _ev(xl3, "'Dashboard Mermas'!D5"),
        'estado_D4': _ev(xl3, "'Dashboard Mermas'!D4"),
    }
    if dash_pobre['estado_D5'] == dash['estado_D5']:
        fallos.append('05:Dashboard Mermas:D5: el estado no se mueve al pasar '
                      'de 8.000 € a 2.000 € de compras (sigue en {!r})'
                      .format(dash_pobre['estado_D5']))
    return {'copia_desechable': dst,
            'analisis_por_categoria': filas,
            'total_mes': total,
            'dashboard_con_8000_de_compras': dash,
            'dashboard_con_2000_de_compras': dash_pobre,
            'dashboard_con_el_registro_vacio': vacio,
            'lectura': 'las dos hojas de explotación estaban vacías de '
                       'fórmulas y el «< 3 % sobre compras» era incalculable: '
                       'no existía ninguna celda donde escribir las compras '
                       'del mes (DOM-23)'}
