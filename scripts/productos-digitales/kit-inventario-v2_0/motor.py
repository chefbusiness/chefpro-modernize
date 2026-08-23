#!/usr/bin/env python3
"""
motor.py — Motor común del Kit de Control de Inventario y Compras v2.0.

Implementa las utilidades TRANSVERSALES del §1 de `kit-inventario-v2-SPEC.md`
sobre los 9 xlsx de `astro-site/public/dl/kit-inventario/`. NO toca ficheros:
recibe un `Workbook` ya cargado y lo modifica en memoria; quien guarda es
`main.py`.

Qué hace (§1):
  1.1  Taxonomía canónica de 10 categorías + 10 unidades como CONSTANTE, con
       validación de datos INLINE (caben en los 255 caracteres de Excel) sobre
       las columnas declaradas en `DV_CATEGORIA` / `DV_UNIDAD`.
  1.2  Formato condicional REAL (`semaforo()` con `containsText`, y reglas
       `expression` para temperatura y variación). Hoy hay 0 reglas en las 30
       hojas: el «semáforo» es el emoji dentro del texto.
  1.3  IVA por categoría en `03!Listas` — tabla, nunca literal en la fórmula.
  1.5  Protección de hoja SIN contraseña con las celdas verdes desbloqueadas,
       y la línea «Revisar → Desproteger hoja (no tiene contraseña)».
  1.7  Formatos por CABECERA (moneda, fecha, porcentaje, entero) — se deducen
       del texto del encabezado, así que sobreviven a que un grupo inserte una
       columna en medio.
  1.10 Bio anclada (es una INSERCIÓN: hoy no la lleva ninguno de los 9), línea
       «Versión 2.0 · agosto 2026 · aichef.pro/kit-inventario · info@aichef.pro»
       y metadata OOXML (`title` / `subject` / `keywords` / `creator`…).

Y expone a los grupos las utilidades que necesitan para el §1.4 (ejemplos),
§1.6 (pestañas), §1.8 (doble guarda) y §1.9 (filas libres):
`insertar_columna`, `insertar_fila`, `expandir_filas`, `guarda_doble`,
`marcar_verde`, `sembrar`, `print_setup`, `linea_instrucciones`, `_reg`.

════════════════════════════════════════════════════════════════════════════
CENTINELAS — por qué el motor no revienta cuando los grupos aún no existen
════════════════════════════════════════════════════════════════════════════
Los rangos que declara la SPEC describen el layout FINAL de la v2.0: `03!C` es
«Categoría» sólo DESPUÉS de que `grupo_b` inserte esa columna (hoy `C8` dice
«Unidad»). Si el motor escribiera la DV de categorías en `03!C` a ciegas, con
`--solo motor` pondría el desplegable de categorías sobre la columna de
unidades — un fichero peor que el de partida.

Por eso CADA objetivo lleva un centinela `(celda, subcadena)`: el motor sólo
actúa si la cabecera ya dice lo que tiene que decir. Lo que no se aplica se
declara en el informe como `pendiente_de_grupo`, con fichero:hoja:celda. Así
`--solo motor` es honesto y el mismo código vale cuando los grupos aterricen.

ORDEN DEL PIPELINE (`main.py`): grupo.pre → motor.aplicar → grupo.post →
motor.cerrar. Todo lo que depende de rangos vive en `cerrar()`, DESPUÉS de que
los grupos hayan insertado columnas y añadido filas. Regla para los grupos:
las columnas se insertan en `pre()`; las filas se AÑADEN (nunca `insertar_fila`
en medio del bloque de datos) en `post()`.

IDEMPOTENTE: todo es escritura ABSOLUTA (siempre el mismo valor en la misma
celda) y los objetos acumulables —validaciones y formato condicional— se
limpian por su marca antes de reescribirse. La 2.ª pasada deja el fichero
byte-equivalente salvo timestamps del zip.

Python 3.7 / openpyxl 3.1.3: sin walrus, sin f-strings de depuración (`{x=}`).
"""
import copy
import re

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Identidad del producto
# ==========================================================================
PID = 'kit-inventario'
CORTO = 'Kit Control de Inventario y Compras'
VERSION = '2.0'

VERSION_LINE = ('Versión 2.0 · agosto 2026 · aichef.pro/kit-inventario · '
                'info@aichef.pro')
RX_VERSION = re.compile(r'^Versi[óo]n \d+\.\d+ · ')

#: §1.10 — bio anclada. OJO: en este kit es una INSERCIÓN, no una sustitución:
#: ninguno de los 9 ficheros lleva bio hoy, así que `postprocess-transversal.py`
#: (que sólo sustituye patrones conocidos) nunca la puso. Literal idéntico al
#: de `postprocess-transversal.py:106` y `kit-pasteleria-v2_0-postprocess.py:77`.
#: No casa `RX_BIO_VIEJA` del censo (`29 años|15 años|años de experiencia|…`):
#: dice «desde los 17 años», que no está en el patrón.
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, '
       'en cocina desde los 17 años · johnguerrero.es')

#: §1.5 — la protección va sin contraseña y hay que decirlo, o el cliente cree
#: que el fichero está cerrado con llave.
NOTA_DESPROTEGER = ('Para editar una celda bloqueada: Revisar → Desproteger '
                    'hoja (no tiene contraseña). Las celdas verdes ya son '
                    'editables sin desproteger nada.')

FICHEROS = [
    '01-inventario-stock-diario.xlsx',
    '02-fichas-proveedores.xlsx',
    '03-pedidos-compra.xlsx',
    '04-recepcion-mercancias.xlsx',
    '05-control-mermas.xlsx',
    '06-fifo-caducidades.xlsx',
    '07-analisis-costes-compras.xlsx',
    'BONUS-08-inventario-rapido-mensual.xlsx',
    'BONUS-09-calculadora-punto-pedido.xlsx',
]

# ==========================================================================
# Paleta y formatos del kit (los de v1.1 + el verde de edición)
# ==========================================================================
VERDE = 'E8F5E9'         # celda EDITABLE por el cliente (convención de familia)
BANDA = 'F5F5F5'         # banda cebra de las filas pares (ya estaba en v1.1)
CAB = '2D2D2D'           # cabecera oscura
CAB_ROJA = 'FF4444'      # cabecera de las hojas de incidencia/merma

CF_VERDE_BG, CF_VERDE_FG = 'C6EFCE', '006100'
CF_AMBAR_BG, CF_AMBAR_FG = 'FFEB9C', '9C6500'
CF_ROJO_BG, CF_ROJO_FG = 'FFC7CE', '9C0006'

FMT_EUR = '#,##0.00 €'
FMT_PCT1 = '0.0%'
FMT_FECHA = 'dd/mm/yyyy'
FMT_ENT = '0'
FMT_CANT = '#,##0.00'

PIE = '© 2026 AI Chef Pro · aichef.pro'

# ==========================================================================
# §1.1 — Taxonomía canónica (LAS 10 de `07!'Coste por Categoría'!A4:A13`)
# ==========================================================================
#: Verificado sobre el fichero real el 2026-08-23:
#: `07-analisis-costes-compras.xlsx:Coste por Categoría:A4` … `:A13`.
#: Es la ÚNICA taxonomía del kit (§7-bis.3). Hoy conviven tres: la de 07 (10),
#: la de `01!C` por rotación cíclica (Pescados/Lácteos/Verduras/Frutas/
#: Secos-Granos/Congelados/Bebidas/Limpieza/Varios/Cárnicos) y la de
#: `05!'Análisis por Categoría'!A4:A10` (7: … Secos, Bebidas).
CATEGORIAS = [
    'Cárnicos',
    'Pescados',
    'Lácteos',
    'Verduras/Frutas',
    'Secos/Granos',
    'Congelados',
    'Bebidas Alcohólicas',
    'Bebidas No Alcohólicas',
    'Limpieza',
    'Otros',
]

#: Unidades REALES de compra en economato (§1.1). «Limpieza» incluye menaje y
#: desechables, de ahí `rollo`, `caja` y `paquete`.
UNIDADES = ['kg', 'L', 'ud', 'docena', 'caja', 'bandeja', 'barril', 'saco',
            'rollo', 'paquete']

#: Alias → canónica. Las categorías sueltas que dejó la v1.1 y que casan 1:1
#: con una de las 10. Lo que NO está aquí se queda como está y lo reasigna el
#: grupo correspondiente (§2/§3): el motor nunca adivina.
ALIAS_CATEGORIA = {
    'verduras': 'Verduras/Frutas',
    'frutas': 'Verduras/Frutas',
    'verduras/frutas': 'Verduras/Frutas',
    'verduras y frutas': 'Verduras/Frutas',
    'secos': 'Secos/Granos',
    'secos/granos': 'Secos/Granos',
    'granos': 'Secos/Granos',
    'bebidas': 'Bebidas No Alcohólicas',
    'varios': 'Otros',
    'otros': 'Otros',
    'carnicos': 'Cárnicos',
    'carne': 'Cárnicos',
    'pescado': 'Pescados',
    'lacteos': 'Lácteos',
    'limpieza y menaje': 'Limpieza',
}

# ==========================================================================
# §1.3 — IVA por categoría
# ==========================================================================
#: Tipos del IVA español aplicables a compras de hostelería (2026). Van a
#: `03!Listas`, NUNCA literales en la fórmula: `03!H` los trae con VLOOKUP y
#: encima lleva una DV `4,10,21` para sobrescribirlos línea a línea (§3).
IVA_POR_CATEGORIA = [
    ('Cárnicos', 10),
    ('Pescados', 10),
    ('Lácteos', 4),
    ('Verduras/Frutas', 4),
    ('Secos/Granos', 10),
    ('Congelados', 10),
    ('Bebidas Alcohólicas', 21),
    ('Bebidas No Alcohólicas', 10),
    ('Limpieza', 21),
    ('Otros', 21),
]

NOTA_IVA = ('Tipo orientativo por categoría: edítalo. El pan común, las '
            'harinas panificables, las legumbres, los cereales, la leche, los '
            'quesos, los huevos, las frutas y las verduras van al 4 %; los '
            'refrescos con azúcares añadidos, al 21 %. En cada línea del '
            'pedido puedes cambiar el tipo con el desplegable 4 / 10 / 21.')

HOJA_LISTAS = 'Listas'

#: Ronda 2 · RD-10/RT-16 — el tipo por CATEGORÍA no puede acertar siempre:
#: dentro de «Secos/Granos» conviven el arroz (4 %) y el aceite (10 %), y
#: dentro de «Bebidas No Alcohólicas» el agua (10 %) y la tónica (21 %). Estos
#: son los productos del kit cuyo tipo NO es el de su categoría; el resto lo
#: hereda. `03!Listas` publica la tabla producto → IVA y `03!'Pedido
#: Actual'!H` busca primero por producto y sólo cae a la categoría si no lo
#: encuentra.
IVA_POR_PRODUCTO = {
    # cereales, harinas y legumbres: superreducido del 4 %
    'Arroz redondo': 4,
    'Pasta seca': 4,
    'Harina de trigo': 4,
    'Legumbres secas': 4,
    # lácteos que NO son leche ni queso: reducido del 10 %
    'Nata 35 % M.G.': 10,
    'Mantequilla': 10,
    # los huevos van al 4 % aunque su categoría en el kit sea «Otros»
    'Huevos M': 4,
    # refrescos con azúcares añadidos: 21 % desde 2024
    'Refresco de cola': 21,
    'Tónica': 21,
}

# ==========================================================================
# Ronda 2 · RD-01/RC-07 — UN SOLO juego de proveedores de ejemplo
# ==========================================================================
#: El kit enviaba DOS directorios ficticios distintos: el 02, el 07 y el
#: BONUS-09 hablaban de «Cárnicas del Norte» y el 03 y el 04 de «Cárnicas del
#: Mercado». Un jefe de compras no puede cotejar el albarán de uno contra la
#: ficha homologada del otro. Ahora los tres grupos importan esta lista.
PROVEEDORES_EJEMPLO = [
    'Cárnicas del Norte',
    'Pescados Ría Fresca',
    'Frutas y Verduras La Huerta',
    'Distribuciones Economato Sur',
    'Bebidas y Distribución Levante',
    'Higiene Profesional HORECA',
]

#: Los mismos nombres con el sufijo que lleva el directorio del 02. Es el
#: literal que tiene que aparecer en el desplegable del 03, en el historial,
#: en la recepción del 04, en el Top 20 del 07 y en la calculadora del
#: BONUS-09: si uno escribe «Cárnicas del Norte» y el desplegable ofrece
#: «Cárnicas del Norte (ejemplo)», filtrar por proveedor no agrupa (RC-07).
PROVEEDORES_MARCADOS = [n + ' (ejemplo)' for n in PROVEEDORES_EJEMPLO]

#: Categoría principal de cada uno, en el MISMO orden.
CATEGORIA_PROVEEDOR = [
    'Cárnicos', 'Pescados', 'Verduras/Frutas', 'Secos/Granos',
    'Bebidas Alcohólicas', 'Limpieza',
]

# ==========================================================================
# Ronda 2 · RD-02/RD-21 — UNA sola tabla de consumo de ejemplo
# ==========================================================================
#: El kit daba TRES puntos de pedido distintos para el mismo producto: el par
#: level del 01, la fórmula de sus Instrucciones y el «punto de pedido» del
#: BONUS-09, que en dos casos superaba el par MÁXIMO del 01 (mandaba reponer
#: por encima de lo que la otra plantilla permite tener). Y el gasto mensual
#: del Top 20 del 07 no cuadraba con ninguno de los dos.
#:
#: A partir de aquí hay UNA definición y las demás se derivan de ella:
#:   par level del 01  = consumo diario × (lead time + cobertura)   ← punto de pedido
#:   gasto mensual 07  = consumo diario × 30 × precio de referencia
#: Los `consumo` están elegidos para que el par level derivado COINCIDA
#: exactamente con el que ya llevaba el 01 (que no se toca).
#: El par MÁXIMO del 01 viaja también aquí: RD-22 · la cantidad sugerida del
#: BONUS-09 estaba capada por vida útil pero no por el sitio donde hay que
#: meter el género, así que sugería 9 barriles de cerveza a un local cuyo par
#: max son 6. Cursar ese pedido significa rechazarlo en la puerta por falta de
#: cámara.
#: producto → (consumo diario, lead time (días), cobertura (días),
#:             vida útil en cámara (días), par MÁXIMO del 01)
CONSUMO_EJEMPLO = {
    'Pechuga de pollo':             (4,    1, 1,   4, 20),
    'Solomillo de ternera':         (1,    2, 1,   8,  8),
    'Salmón fresco':                (2,    1, 1,   3, 10),
    'Gambas':                       (1.5,  1, 1,   3,  8),
    'Leche entera':                 (8,    2, 1,  12, 60),
    'Patata':                       (6.25, 2, 2,  30, 60),
    'Aceite de oliva virgen extra': (2,    4, 1, 540, 25),
    'Arroz redondo':                (2,    4, 1, 720, 25),
    'Queso parmesano':              (0.4,  4, 1, 120,  5),
    'Cerveza de grifo (barril 30 L)': (0.25, 3, 5, 120, 6),
}

# ==========================================================================
# Registro de fórmulas (main.py verifica una a una que quedaron con cache)
# ==========================================================================
REGISTRO = []


def _reg(ws, coord, formula):
    """Anota una fórmula escrita para que `main.py` la verifique con
    `data_only`. Los grupos DEBEN llamarlo por cada fórmula que escriban."""
    REGISTRO.append((ws.title, coord, formula))
    return formula


# ==========================================================================
# Bloques de datos por hoja: (hoja, fila_cabecera, primera_fila, última_fila)
# ==========================================================================
#: `(hoja, fila_cabecera, primera_fila, ULTIMA_HOY, ULTIMA_V2)`.
#:
#: Hacen falta las DOS últimas filas, no basta con recortar a `ws.max_row`:
#: por debajo del bloque de datos hay COLA (el «TOTAL MES:» de `05` en la 55,
#: el pie «© 2026 AI Chef Pro» de `01!Cocina` en la 27, el SUBTOTAL/TOTAL de
#: `03` en las 30-31). Con `min(ULTIMA_V2, ws.max_row)` la DV de categorías
#: caía sobre la fila del total y el verde de edición sobre el pie de página.
#: Regla: si `ws.max_row` ya supera `ULTIMA_V2`, el grupo expandió el bloque y
#: se usa `ULTIMA_V2`; si no, el bloque sigue siendo el de la v1.1 y se usa
#: `ULTIMA_HOY`.
#:
#: Se omite a propósito `06!'Alertas Caducidad'`: hoy es texto en la columna A,
#: no una tabla (grupo_c le pone el autofiltro y los contadores).
BLOQUES = {
    '01-inventario-stock-diario.xlsx': [
        ('Cocina', 4, 5, 24, 44),
        ('Barra', 4, 5, 19, 34),
        ('Almacén', 4, 5, 19, 34),
        ('Resumen Dashboard', 3, 4, 3, 20),
    ],
    '02-fichas-proveedores.xlsx': [
        ('Directorio Proveedores', 3, 4, 23, 23),
        ('Comparativa Precios', 3, 4, 13, 13),
        ('Evaluación Proveedores', 3, 4, 13, 13),
        ('Condiciones Comerciales', 3, 4, 13, 13),
    ],
    '03-pedidos-compra.xlsx': [
        ('Pedido Actual', 8, 9, 28, 38),
        ('Historial Pedidos', 3, 4, 3, 43),
        ('Proveedores', 3, 4, 23, 23),
    ],
    '04-recepcion-mercancias.xlsx': [
        ('Control Recepción', 3, 4, 33, 43),
        ('Registro Incidencias', 3, 4, 3, 23),
        # Ronda 2 · RD-23 — la tabla legal pasa de 13 a 15 familias (faltaban
        # las bebidas y el material no alimentario) y gana la columna de
        # mínimo, así que el bloque llega ahora a la fila 18.
        ('Verificación Temperaturas', 3, 4, 18, 18),
    ],
    '05-control-mermas.xlsx': [
        ('Registro Diario Mermas', 3, 4, 53, 103),
        ('Análisis por Categoría', 3, 4, 10, 13),
        ('Dashboard Mermas', 3, 4, 8, 17),
        ('Plan de Acción', 3, 4, 3, 23),
    ],
    '06-fifo-caducidades.xlsx': [
        ('Control FIFO', 4, 5, 54, 54),
        # Ronda 2 · RD-04/RD-05 — el mapa pasa de 9 a 11 zonas (huevera y
        # almacén de químicos), así que el bloque llega a la fila 14.
        ('Mapa Almacén', 3, 4, 14, 14),
    ],
    '07-analisis-costes-compras.xlsx': [
        ('Coste por Categoría', 3, 4, 13, 13),
        ('Top 20 Productos', 3, 4, 23, 23),
        ('Dashboard KPIs', 3, 4, 9, 17),
        ('Evolución Mensual', 3, 4, 15, 15),
    ],
    'BONUS-08-inventario-rapido-mensual.xlsx': [
        ('Conteo Rápido', 4, 5, 54, 84),
    ],
    'BONUS-09-calculadora-punto-pedido.xlsx': [
        ('Calculadora', 3, 4, 33, 33),
        # Ronda 2 · RT-08 — «Factor de vida útil» pasa a ser un parámetro de
        # verdad (D12) y se añade la fila 13 del stock máximo (RD-22).
        ('Parámetros', 3, 4, 13, 13),
    ],
}

#: Hojas de REFERENCIA o de PARÁMETROS: el verde no se deduce solo (lo pondría
#: sobre la base normativa o sobre el rótulo de un parámetro). Los grupos
#: marcan ahí las celdas concretas con `marcar_verde()`; el motor no toca.
SIN_VERDE_AUTO = frozenset({
    'Verificación Temperaturas',   # 04: umbrales legales, tabla de consulta
    'Parámetros',                  # BONUS-09: D4/D5 los marca grupo_c
    HOJA_LISTAS,                   # 03: se deja la hoja entera sin proteger
})

#: Ronda 2 · RT-13/RC-14 — hojas que hay que proteger AUNQUE no tengan ni una
#: celda verde. `proteger()` deja abierta la hoja sin verdes para no entregar
#: un libro donde el cliente no puede escribir nada; pero estas cinco son
#: justo lo contrario: tres son 100 % fórmulas y su propia línea 2 dice «aquí
#: no se escribe nada», y las otras dos son tablas de referencia de las que
#: dependen los VLOOKUP de sus hermanas protegidas (un clic borraba la tabla
#: legal de temperaturas del 04 sin un solo aviso).
PROTEGER_SIN_VERDE = frozenset({
    'Resumen Dashboard',           # 01: 33 fórmulas, 0 entradas
    'Alertas Caducidad',           # 06: 12 fórmulas, 0 entradas
    'Evolución Mensual',           # 07: 35 fórmulas, 0 entradas
    'Verificación Temperaturas',   # 04: tabla legal que alimenta el VLOOKUP
})

#: Cabeceras que NUNCA llevan verde aunque su columna no tenga fórmulas.
NO_VERDE_CABECERAS = frozenset({
    'base normativa', 'cómo calcular', 'como calcular', 'categorías',
    'kpi', 'parámetro', 'parametro',
})


# ==========================================================================
# §1.1 — Validación de datos INLINE
# ==========================================================================
#: `(hoja, rango_col, centinela_celda, centinela_texto)`. El rango de filas sale
#: de `BLOQUES`. Referencias de la SPEC §1.1: `01!C/D`, `03!C/D`,
#: `05!'Registro Diario Mermas'!C/E`, `06!'Control FIFO'!C`, `07!'Top 20'!C`
#: y `BONUS-08!C/D`.
DV_CATEGORIA = {
    '01-inventario-stock-diario.xlsx': [
        ('Cocina', 'C', 'C4', 'Categoría'),
        ('Barra', 'C', 'C4', 'Categoría'),
        ('Almacén', 'C', 'C4', 'Categoría'),
    ],
    '03-pedidos-compra.xlsx': [
        ('Pedido Actual', 'C', 'C8', 'Categoría'),
    ],
    '05-control-mermas.xlsx': [
        ('Registro Diario Mermas', 'C', 'C3', 'Categoría'),
    ],
    '06-fifo-caducidades.xlsx': [
        ('Control FIFO', 'C', 'C4', 'Categoría'),
    ],
    '07-analisis-costes-compras.xlsx': [
        ('Top 20 Productos', 'C', 'C3', 'Categoría'),
    ],
    'BONUS-08-inventario-rapido-mensual.xlsx': [
        ('Conteo Rápido', 'C', 'C4', 'Categoría'),
    ],
}

DV_UNIDAD = {
    '01-inventario-stock-diario.xlsx': [
        ('Cocina', 'D', 'D4', 'Unidad'),
        ('Barra', 'D', 'D4', 'Unidad'),
        ('Almacén', 'D', 'D4', 'Unidad'),
    ],
    '03-pedidos-compra.xlsx': [
        ('Pedido Actual', 'D', 'D8', 'Unidad'),
    ],
    '05-control-mermas.xlsx': [
        ('Registro Diario Mermas', 'E', 'E3', 'Unidad'),
    ],
    'BONUS-08-inventario-rapido-mensual.xlsx': [
        ('Conteo Rápido', 'D', 'D4', 'Unidad'),
    ],
}

#: Marca que llevan las DV del motor en `promptTitle`, para poder limpiarlas
#: antes de reescribirlas (si no, cada pasada acumularía una copia y la
#: idempotencia daría diferencias en `dv`).
MARCA_DV = 'kitinv-v2'


# ==========================================================================
# §1.2 — Formato condicional
# ==========================================================================
#: Vocabulario del semáforo: (subcadena, color). Se evalúa EN ORDEN y cada
#: regla lleva `stopIfTrue`, así que lo más grave gana. `containsText` usa
#: SEARCH, que no distingue mayúsculas ni se pelea con el emoji de delante.
VOC_STOCK = [('PEDIR', 'rojo'), ('BAJO', 'ambar'), ('OK', 'verde')]
VOC_FIFO = [('CADUCADO', 'rojo'), ('URGENTE', 'rojo'),
            ('REVISAR', 'ambar'), ('PRÓXIMO', 'ambar'), ('OK', 'verde')]
#: Ronda 2 · RT-01/RD-06/RT-15 — el veredicto de temperatura tiene ahora
#: CINCO salidas: rechazo por calor, rechazo por frío (un pescado a -20 °C
#: venía congelado), «familia sin límite» cuando la familia no resuelve en la
#: tabla legal —antes eso daba «CONFORME» a cualquier temperatura— y N/A.
VOC_RECEPCION = [('RECHAZAR', 'rojo'), ('SIN LÍMITE', 'ambar'),
                 ('CONFORME', 'verde')]
VOC_KPI = [('ALERTA', 'rojo'), ('REVISAR', 'ambar'), ('OK', 'verde')]
#: Ronda 2 · RD-18 — estado de la cotización de `02!'Comparativa Precios'!U`.
#: OJO al orden y a las subcadenas: «VENCIDA» no contiene «vence».
VOC_COTIZACION = [('VENCIDA', 'rojo'), ('vence', 'ambar'),
                  ('vigente', 'verde')]

#: `(hoja, columna, centinela_celda, centinela_texto, vocabulario)`.
CF_SEMAFORO = {
    '01-inventario-stock-diario.xlsx': [
        ('Cocina', 'H', 'H4', 'Estado', VOC_STOCK),
        ('Barra', 'H', 'H4', 'Estado', VOC_STOCK),
        ('Almacén', 'H', 'H4', 'Estado', VOC_STOCK),
    ],
    '02-fichas-proveedores.xlsx': [
        ('Comparativa Precios', 'U', 'U3', 'Estado', VOC_COTIZACION),
    ],
    '04-recepcion-mercancias.xlsx': [
        # Ronda 2 · el layout crece a A:V (categoría del kit, familia
        # sugerida, precio/ud, valor de la diferencia y Tª mín.), así que el
        # veredicto pasa de M a R.
        ('Control Recepción', 'R', 'R3', 'Conform', VOC_RECEPCION),
    ],
    '05-control-mermas.xlsx': [
        ('Dashboard Mermas', 'D', 'D3', 'Estado', VOC_KPI),
    ],
    '06-fifo-caducidades.xlsx': [
        ('Control FIFO', 'L', 'L4', 'Estado', VOC_FIFO),
    ],
    '07-analisis-costes-compras.xlsx': [
        ('Dashboard KPIs', 'E', 'E3', 'Estado', VOC_KPI),
    ],
}

#: Reglas `expression` (§1.2): la temperatura del 04 y la variación del 07 son
#: NÚMEROS, así que no hay texto que buscar — el color sale de una condición.
#: `{f}` se sustituye por la primera fila del rango (referencia relativa).
CF_EXPRESION = {
    '04-recepcion-mercancias.xlsx': [
        # `="✗ RECHAZAR"` ya no vale: el veredicto distingue el rechazo por
        # CALOR del rechazo por FRÍO, así que se busca la subcadena.
        ('Control Recepción', 'O', 'O3', 'Temp',
         '=AND($O{f}<>"",ISNUMBER(SEARCH("RECHAZAR",$R{f})))', 'rojo'),
    ],
    '07-analisis-costes-compras.xlsx': [
        ('Top 20 Productos', 'J', 'J3', 'Variaci',
         '=AND($J{f}<>"",$J{f}>0.05)', 'rojo'),
    ],
}


# ==========================================================================
# §1.7 — Formatos deducidos de la CABECERA
# ==========================================================================
#: Se decide por el TEXTO del encabezado, no por la letra de la columna: así
#: sobrevive a que `grupo_b` inserte «Categoría» delante de «Unidad» en el 03.
#: El orden importa — se para en la primera coincidencia.
FORMATO_POR_CABECERA = [
    (FMT_PCT1, ('%',)),
    (FMT_FECHA, ('fecha', 'caducidad', 'vigencia', 'homologación',
                 'homologacion')),
    (FMT_ENT, ('días', 'dias', 'lead time', 'nº incidencias',
               'n° incidencias', 'cubiertos')),
    (FMT_EUR, ('€', 'eur', 'precio', 'coste', 'importe', 'subtotal',
               'valor', 'gasto', 'abono', 'total', 'ventas', 'compras',
               'ticket', 'rappel')),
]

#: Rangos de importe que NINGUNA cabecera delata: los 12 meses del 07 se
#: llaman «Ene»…«Dic» (§1.7 pide `#,##0.00 €` en `07!B4:M13`).
RANGOS_FORMATO = {
    '07-analisis-costes-compras.xlsx': [
        ('Coste por Categoría', 'B4:N14', FMT_EUR, 'A3', 'Categoría'),
        ('Evolución Mensual', 'B4:B15', FMT_EUR, 'A3', 'Mes'),
    ],
}


# ==========================================================================
# Utilidades de estilo
# ==========================================================================
def _relleno(cel):
    f = cel.fill
    if f is None or f.fill_type != 'solid' or f.fgColor is None:
        return None
    rgb = f.fgColor.rgb
    return rgb.upper() if isinstance(rgb, str) else None


def es_verde(cel):
    """True si la celda lleva el verde de edición del kit."""
    r = _relleno(cel)
    return bool(r) and r.endswith(VERDE)


def marcar_verde(ws, rango):
    """Pinta de verde (= editable) un rango `A1:B9` o una celda `A1`."""
    n = 0
    filas = ws[rango]
    # `ws['C4']` devuelve un Cell suelto (no iterable) y `ws['C4:D9']` una
    # tupla de tuplas: sin esto, marcar UNA celda reventaba con «'Cell' object
    # is not iterable» pese a lo que promete la docstring.
    if not isinstance(filas, tuple):
        filas = ((filas,),)
    for fila in filas:
        celdas = fila if isinstance(fila, tuple) else (fila,)
        for c in celdas:
            if not es_verde(c):
                c.fill = PatternFill('solid', fgColor=VERDE)
            c.protection = Protection(locked=False)
            n += 1
    return n


def sembrar(ws, fila, valores, col0=1, marca=None):
    """§1.4 — escribe una fila de ejemplo. `valores` es una lista; `None` deja
    la celda como esté (útil para no pisar una fórmula). `marca` se concatena
    a la primera celda de texto ('(ejemplo)')."""
    escritas = 0
    for i, v in enumerate(valores):
        if v is None:
            continue
        cel = ws.cell(row=fila, column=col0 + i)
        if marca and isinstance(v, str) and escritas == 0:
            v = v + ' ' + marca
        cel.value = v
        escritas += 1
    return escritas


def guarda_doble(col_cant, col_precio, fila, aviso='⚠ falta coste'):
    """§1.8 — doble guarda de un producto cantidad × precio.

    Las guardas de la v1.1 vigilan SÓLO la cantidad (`=IF(D4="","",D4*F4)`), así
    que una línea con cantidad y sin precio vale 0,00 € y se suma al total como
    si fuera gratis. Devuelve
    `=IF($D4="","",IF($F4="",\"⚠ falta coste\",$D4*$F4))`.
    """
    return ('=IF(${c}{f}="","",IF(${p}{f}="","{a}",${c}{f}*${p}{f}))'
            .format(c=col_cant, p=col_precio, f=fila, a=aviso))


# ==========================================================================
# Utilidades genéricas (probadas en kit-pasteleria-v2_0 y kit-escandallos-v2_0)
# ==========================================================================
RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _traducir_formula(valor, idx, eje):
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        ci = column_index_from_string(col)
        fi = int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        return '{}{}{}{}'.format(d1, col, d2, fila)

    return RX_REF.sub(_sub, valor)


def _rangos_dv(ws):
    return [({k: getattr(dv, k, None) for k in CAMPOS_DV},
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**{k: v for k, v in attrs.items() if v is not None})
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje) if idx else r)


def _desplazar_rango(ref, idx, eje):
    partes = ref.split(':')
    fuera = []
    for p in partes:
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        fuera.append('{}{}{}{}'.format(d1, col, d2, fila))
    return ':'.join(fuera)


def insertar_columna(ws, idx):
    """Inserta una columna en `idx` manteniendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos de columna.

    Los grupos la llaman desde `pre()`, ANTES de que el motor escriba nada que
    dependa de rangos.
    """
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos = dict((k, v.width) for k, v in ws.column_dimensions.items()
                  if v.width)

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra, ancho in sorted(anchos.items(),
                               key=lambda kv: -column_index_from_string(kv[0])):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = ancho


def insertar_fila(ws, idx):
    """Equivalente por filas de `insertar_columna`."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = dict((k, v.height) for k, v in ws.row_dimensions.items()
                   if v.height)

    for fila in range(max_r, idx - 1, -1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila + 1, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'fila'))
    _restaurar_dv(ws, dvs, idx, 'fila')

    for fila, alto in sorted(alturas.items(), reverse=True):
        if fila >= idx:
            ws.row_dimensions[fila + 1].height = alto


def expandir_filas(ws, r_ult, r_destino, cola=(), numerar=None):
    """§1.9 — añade filas libres AL FINAL del bloque de datos replicando la
    última fila (estilo, fórmulas traducidas, validaciones ya cubiertas por su
    rango) y baja la «cola» (totales, pie) `r_destino - r_ult` filas.

    `cola` = filas por debajo del bloque que hay que mover, de arriba abajo.
    `numerar` = índice de columna (1 = A) con el número de línea, o None.

    CENTINELA de idempotencia: si `r_destino` ya tiene contenido replicado, no
    hace nada. Devuelve el nº de filas añadidas.
    """
    if r_destino <= r_ult:
        return 0
    max_c = ws.max_column
    if any(ws.cell(row=r_destino, column=c).value is not None
           for c in range(1, max_c + 1)):
        return 0                                    # ya expandida
    delta = r_destino - r_ult

    # 1) mover la cola hacia abajo, de abajo arriba para no pisarse.
    #    Las combinaciones se DESHACEN todas primero y se rehacen al final: si
    #    se rehace la del destino antes de copiar, las celdas B..H de la fila
    #    de llegada ya son `MergedCell` y `dst.value = …` revienta con
    #    «object attribute 'value' is read-only» (medido con el pie
    #    «© 2026 AI Chef Pro», combinado A27:H27, de `01!Cocina`).
    rehacer = []
    for m in [str(r) for r in ws.merged_cells.ranges]:
        partes = m.split(':')
        mm = RX_REF.fullmatch(partes[0])
        if mm and int(mm.group(4)) in cola:
            ws.unmerge_cells(m)
            rehacer.append(':'.join(_corre_ref(p, r_ult, delta)
                                    for p in partes))
    for origen in sorted(cola, reverse=True):
        for c in range(1, max_c + 1):
            src = ws.cell(row=origen, column=c)
            dst = ws.cell(row=origen + delta, column=c)
            dst.value = _corre_cola(src.value, r_ult, r_destino, delta)
            dst._style = copy.copy(src._style)
            src.value = None
    for m in rehacer:
        ws.merge_cells(m)

    # 2) replicar la última fila del bloque hacia abajo. El ESTILO se toma de
    #    `fila - 2` para que la banda cebra siga alternando (copiar siempre de
    #    `r_ult` dejaría 50 filas del mismo color); la FÓRMULA sale de `r_ult`
    #    traducida a la fila nueva.
    base = ws.cell(row=r_ult, column=numerar).value if numerar else None
    for fila in range(r_ult + 1, r_destino + 1):
        plantilla = fila - 2 if fila - 2 >= 1 else r_ult
        for c in range(1, max_c + 1):
            src = ws.cell(row=r_ult, column=c)
            dst = ws.cell(row=fila, column=c)
            dst._style = copy.copy(ws.cell(row=plantilla, column=c)._style)
            if isinstance(src.value, str) and src.value.startswith('='):
                dst.value = _corre_filas(src.value, 0, fila - r_ult)
                _reg(ws, dst.coordinate, dst.value)
            elif numerar and c == numerar and isinstance(base, int):
                dst.value = base + (fila - r_ult)
            else:
                dst.value = None
    return delta


def _corre_ref(ref, desde, delta):
    m = RX_REF.fullmatch(ref)
    if not m:
        return ref
    d1, col, d2, fila = m.groups()
    fi = int(fila)
    return '{}{}{}{}'.format(d1, col, d2, fi + delta if fi > desde else fi)


def _corre_cola(valor, r_ult, r_destino, delta):
    """Traduce una fórmula de la COLA (totales, pie) al expandir el bloque.

    Dos reglas, no una:
      * una referencia a una fila POR DEBAJO del bloque se desplaza `delta`
        (el `F30`/`F31` de los rótulos SUBTOTAL/TOTAL del 03);
      * una referencia a la ÚLTIMA fila del bloque se estira hasta la nueva
        última fila. Sin esto, `05!G55 = SUM(G4:G53)` seguiría sumando 50
        filas después de crecer a 100: las 50 mermas nuevas no entrarían en el
        total y el «TOTAL MES» mentiría (§1.9, «su SUM cubre el rango
        completo»).
    """
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        fi = int(fila)
        if fi == r_ult:
            fila = str(r_destino)
        elif fi > r_ult:
            fila = str(fi + delta)
        return '{}{}{}{}'.format(d1, col, d2, fila)

    return RX_REF.sub(_sub, valor)


def _corre_filas(valor, desde, delta):
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        fi = int(fila)
        if fi > desde:
            fila = str(fi + delta)
        return '{}{}{}{}'.format(d1, col, d2, fila)

    return RX_REF.sub(_sub, valor)


def print_setup(ws, header_row=None, landscape=True):
    """A4, ajustado al ancho, con pie de página. El censo cuenta como defecto
    (`noprint`) toda hoja con `paperSize != 9`."""
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = '{0}:{0}'.format(header_row)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def linea_instrucciones(ws, texto, rx=None, col=1):
    """Escribe `texto` en Instrucciones: sustituye la línea que case con `rx` o
    la añade al final si no existe. Nunca duplica."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return r
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return r
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)
    return destino


# ==========================================================================
# Centinelas
# ==========================================================================
def _centinela(ws, celda, texto):
    """True si la cabecera ya dice lo que tiene que decir. Ver la cabecera del
    módulo: es lo que impide que el motor escriba la DV de categorías sobre la
    columna de unidades cuando el grupo aún no ha insertado la suya."""
    try:
        v = ws[celda].value
    except (ValueError, KeyError):
        return False
    return isinstance(v, str) and texto.lower() in v.lower()


def _bloque(fname, hoja):
    for entrada in BLOQUES.get(fname, []):
        if entrada[0] == hoja:
            return entrada[1], entrada[2], entrada[3], entrada[4]
    return None


def _rango_filas(ws, fname, hoja):
    """Devuelve (r0, r1) del bloque de DATOS, sin tragarse la cola.

    `ws.max_row` NO sirve como tope: incluye el «TOTAL MES:» y el pie
    «© 2026 AI Chef Pro». Si la hoja ya creció por encima de `ULTIMA_V2` es
    que un grupo la expandió (§1.9) y el bloque llega hasta ahí; si no, sigue
    siendo el de la v1.1.
    """
    b = _bloque(fname, hoja)
    if not b:
        return None
    hdr, r0, r1_hoy, r1_v2 = b
    r1 = r1_v2 if ws.max_row > r1_v2 else min(r1_hoy, ws.max_row)
    if r1 < r0:
        return None
    return r0, r1


# ==========================================================================
# §1.1 — aplicar las validaciones
# ==========================================================================
def _dv_inline(valores, titulo, prompt):
    """DV de lista INLINE. Excel limita `formula1` a 255 caracteres contando
    las comillas: las 10 categorías ocupan ~130 y las 10 unidades ~55, así que
    caben sin hoja auxiliar (§1.1). Si algún día no cupieran, esto ABORTA en
    vez de escribir una lista truncada."""
    formula = '"{}"'.format(','.join(valores))
    if len(formula) > 255:
        raise ValueError('DV inline de {} caracteres (>255): {}'
                         .format(len(formula), titulo))
    return DataValidation(
        type='list', formula1=formula, allow_blank=True,
        showErrorMessage=True,
        errorTitle=titulo,
        error='Elige un valor de la lista. Es la taxonomía única del kit: '
              'si cada plantilla usa palabras distintas, los SUMIF de los '
              'análisis no suman nada.',
        errorStyle='stop',
        showInputMessage=True,
        promptTitle='{} · {}'.format(MARCA_DV, titulo),
        prompt=prompt)


def _limpiar_dv(ws):
    """Quita las DV que puso el motor en pasadas anteriores (las reconoce por
    `promptTitle`). Sin esto la 2.ª pasada acumularía una copia y la
    idempotencia daría `cambia dv`."""
    vivos = []
    for dv in ws.data_validations.dataValidation:
        pt = getattr(dv, 'promptTitle', None) or ''
        if not pt.startswith(MARCA_DV):
            vivos.append(dv)
    ws.data_validations.dataValidation = vivos


def aplicar_dv(wb, fname, informe, pendientes):
    """§1.1 — DV de categorías y unidades sobre las columnas declaradas."""
    puestas = 0
    for tabla, valores, titulo, prompt in (
            (DV_CATEGORIA, CATEGORIAS, 'Categoría no válida',
             'Las 10 categorías del kit. Son las mismas en las 9 plantillas: '
             'los análisis del 05 y del 07 agregan por este texto.'),
            (DV_UNIDAD, UNIDADES, 'Unidad no válida',
             'Unidad REAL de compra: la cerveza de grifo va en barril, los '
             'huevos en docena y el vino en ud (botella de 75 cl).')):
        for hoja, col, cent_cel, cent_txt in tabla.get(fname, []):
            if hoja not in wb.sheetnames:
                pendientes.append('{}:{}: la hoja no existe todavía'
                                  .format(fname, hoja))
                continue
            ws = wb[hoja]
            if not _centinela(ws, cent_cel, cent_txt):
                pendientes.append(
                    '{}:{}:{}: la cabecera no dice «{}» todavía — DV de {} '
                    'aplazada al grupo'.format(fname, hoja, cent_cel, cent_txt,
                                               titulo.split()[0].lower()))
                continue
            rango = _rango_filas(ws, fname, hoja)
            if not rango:
                pendientes.append('{}:{}: sin bloque de datos'
                                  .format(fname, hoja))
                continue
            dv = _dv_inline(valores, titulo, prompt)
            ws.add_data_validation(dv)
            ref = '{c}{a}:{c}{b}'.format(c=col, a=rango[0], b=rango[1])
            dv.add(ref)
            puestas += 1
            informe.append('{}:{}!{}: DV inline ({} valores)'
                           .format(fname, hoja, ref, len(valores)))
    return puestas


# ==========================================================================
# §1.2 — formato condicional REAL
# ==========================================================================
def _dxf(color):
    bg, fg = {'verde': (CF_VERDE_BG, CF_VERDE_FG),
              'ambar': (CF_AMBAR_BG, CF_AMBAR_FG),
              'rojo': (CF_ROJO_BG, CF_ROJO_FG)}[color]
    return DifferentialStyle(font=Font(color=fg, bold=True),
                             fill=PatternFill(start_color=bg, end_color=bg,
                                              fill_type='solid'))


def _limpiar_cf(ws, refs):
    """Borra las reglas de los rangos que gobierna el motor y deja intactas las
    de los grupos. `refs` = conjunto de sqref en texto."""
    nueva = ConditionalFormattingList()
    for cf in ws.conditional_formatting:
        if str(cf.sqref) in refs:
            continue
        for regla in cf.rules:
            nueva.add(str(cf.sqref), regla)
    ws.conditional_formatting = nueva


def semaforo(ws, rango, vocabulario):
    """§1.2 — pinta un rango de TEXTO según las palabras del vocabulario.

    Hoy las 30 hojas tienen 0 reglas de formato condicional: el «semáforo» que
    prometen las Instrucciones es el emoji dentro de la cadena. Esto lo hace
    real: la celda se colorea, y se ve igual en una impresión en blanco y negro
    porque el emoji sigue ahí.
    """
    primera = int(RX_REF.fullmatch(rango.split(':')[0]).group(4))
    col = RX_REF.fullmatch(rango.split(':')[0]).group(2)
    for texto, color in vocabulario:
        formula = 'NOT(ISERROR(SEARCH("{t}",{c}{f})))'.format(
            t=texto, c=col, f=primera)
        ws.conditional_formatting.add(rango, Rule(
            type='containsText', operator='containsText', text=texto,
            dxf=_dxf(color), stopIfTrue=True, formula=[formula]))
    return len(vocabulario)


def aplicar_cf(wb, fname, informe, pendientes):
    """§1.2 — semáforos por texto + las dos reglas de expresión."""
    puestas = 0
    porhoja = {}
    for hoja, col, cc, ct, voc in CF_SEMAFORO.get(fname, []):
        porhoja.setdefault(hoja, []).append(('texto', col, cc, ct, voc, None))
    for hoja, col, cc, ct, expr, color in CF_EXPRESION.get(fname, []):
        porhoja.setdefault(hoja, []).append(('expr', col, cc, ct, expr, color))

    for hoja, objetivos in porhoja.items():
        if hoja not in wb.sheetnames:
            pendientes.append('{}:{}: la hoja no existe todavía'
                              .format(fname, hoja))
            continue
        ws = wb[hoja]
        rango_filas = _rango_filas(ws, fname, hoja)
        if not rango_filas:
            pendientes.append('{}:{}: sin bloque de datos'.format(fname, hoja))
            continue
        refs = set()
        listos = []
        for tipo, col, cc, ct, dato, color in objetivos:
            if not _centinela(ws, cc, ct):
                pendientes.append(
                    '{}:{}:{}: la cabecera no dice «{}» todavía — formato '
                    'condicional aplazado al grupo'.format(fname, hoja, cc, ct))
                continue
            ref = '{c}{a}:{c}{b}'.format(c=col, a=rango_filas[0],
                                         b=rango_filas[1])
            refs.add(ref)
            listos.append((tipo, ref, col, rango_filas[0], dato, color))
        _limpiar_cf(ws, refs)
        for tipo, ref, col, f0, dato, color in listos:
            if tipo == 'texto':
                n = semaforo(ws, ref, dato)
            else:
                ws.conditional_formatting.add(ref, Rule(
                    type='expression', dxf=_dxf(color), stopIfTrue=True,
                    formula=[dato.format(f=f0).lstrip('=')]))
                n = 1
            puestas += n
            informe.append('{}:{}!{}: {} regla(s) de formato condicional'
                           .format(fname, hoja, ref, n))
    return puestas


# ==========================================================================
# §1.3 — tabla de IVA por categoría (03!Listas)
# ==========================================================================
def tabla_iva(wb, informe):
    """§1.3 — crea/rellena `Listas` con categoría → IVA %.

    El 03 tiene hoy el `10` escrito a mano en las 20 líneas del pedido
    (`Pedido Actual!G9:G28`), así que un pedido de vino sale con un 11 % de IVA
    de menos. La tabla es el origen del `VLOOKUP` de `03!H` (§3) y el cliente
    puede editarla: sus celdas de tipo van en VERDE. Ronda 2 (RT-13): la hoja
    SÍ se protege —como el resto—, porque dejarla abierta permitía borrar de
    un clic la tabla de la que depende una hoja protegida.
    """
    if HOJA_LISTAS in wb.sheetnames:
        ws = wb[HOJA_LISTAS]
    else:
        ws = wb.create_sheet(HOJA_LISTAS)
        informe.append('03:{}: hoja creada (§1.3)'.format(HOJA_LISTAS))
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 70
    ws['A1'] = 'Categoría'
    ws['B1'] = 'IVA %'
    ws['C1'] = 'Nota'
    for cel in (ws['A1'], ws['B1'], ws['C1']):
        cel.fill = PatternFill('solid', fgColor=CAB)
        cel.font = Font(bold=True, color='FFFFFF')
    for i, par in enumerate(IVA_POR_CATEGORIA):
        fila = 2 + i
        ws.cell(row=fila, column=1, value=par[0])
        cel = ws.cell(row=fila, column=2, value=par[1])
        cel.number_format = FMT_ENT
        cel.fill = PatternFill('solid', fgColor=VERDE)
        cel.protection = Protection(locked=False)
    ws['C2'] = NOTA_IVA
    ws['C2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('C2:C11')

    # ---- Ronda 2 · RD-10/RT-16 · segunda tabla, producto → IVA ----------
    # Ninguna asignación POR CATEGORÍA puede acertar: en «Secos/Granos»
    # conviven el arroz (4 %) y el aceite (10 %), y en «Bebidas No
    # Alcohólicas» el agua (10 %) y la tónica (21 %). Sólo se listan los
    # productos cuyo tipo NO coincide con el de su categoría: los demás ya
    # salen bien por la tabla de arriba y duplicarlos sería una segunda
    # verdad que mantener.
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 62
    ws['E1'] = 'Producto'
    ws['F1'] = 'IVA %'
    ws['G1'] = 'Por qué no es el de su categoría'
    for cel in (ws['E1'], ws['F1'], ws['G1']):
        cel.fill = PatternFill('solid', fgColor=CAB)
        cel.font = Font(bold=True, color='FFFFFF')
    for i, nombre in enumerate(sorted(IVA_POR_PRODUCTO)):
        fila = 2 + i
        ws.cell(row=fila, column=5, value=nombre)
        cel = ws.cell(row=fila, column=6, value=IVA_POR_PRODUCTO[nombre])
        cel.number_format = FMT_ENT
        cel.fill = PatternFill('solid', fgColor=VERDE)
        cel.protection = Protection(locked=False)
        ws.cell(row=fila, column=5).fill = PatternFill('solid',
                                                       fgColor=VERDE)
        ws.cell(row=fila, column=5).protection = Protection(locked=False)
    ws['G2'] = (
        'EXCEPCIONES por producto. El pedido busca PRIMERO el nombre del '
        'producto en esta tabla y sólo si no lo encuentra usa el tipo de su '
        'categoría. Aquí están únicamente los productos del kit cuyo tipo NO '
        'es el de su categoría: los cereales, las harinas y las legumbres '
        'van al 4 % aunque su categoría (Secos/Granos) sea del 10 %; la nata '
        'y la mantequilla al 10 % aunque Lácteos sea del 4 %; los refrescos '
        'con azúcares añadidos al 21 %. Añade los tuyos escribiendo el '
        'nombre EXACTO que uses en el pedido.')
    ws['G2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('G2:G{}'.format(1 + max(len(IVA_POR_PRODUCTO), 2)))
    print_setup(ws, 1, landscape=False)
    return len(IVA_POR_CATEGORIA)


# ==========================================================================
# §1.1 — normalización de las categorías que YA casan con una canónica
# ==========================================================================
def normalizar_categorias(wb, fname, informe):
    """Reescribe las categorías sueltas de la v1.1 que casan 1:1 con una de las
    10 (`Secos`→`Secos/Granos`, `Bebidas`→`Bebidas No Alcohólicas`,
    `Varios`→`Otros`…). Lo que no está en `ALIAS_CATEGORIA` NO se toca: la
    reasignación por producto es trabajo de los grupos (§2/§3), no del motor.
    """
    cambiadas = 0
    objetivos = []
    for hoja, col, cc, ct in DV_CATEGORIA.get(fname, []):
        objetivos.append((hoja, col, cc, ct))
    # `05!'Análisis por Categoría'!A` y `07!'Coste por Categoría'!A` son la
    # taxonomía impresa, no un campo de entrada: también se normalizan.
    objetivos += {
        '05-control-mermas.xlsx': [('Análisis por Categoría', 'A', 'A3',
                                    'Categoría')],
        '07-analisis-costes-compras.xlsx': [('Coste por Categoría', 'A', 'A3',
                                             'Categoría')],
    }.get(fname, [])

    canon = dict((c.lower(), c) for c in CATEGORIAS)
    for hoja, col, cc, ct in objetivos:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        if not _centinela(ws, cc, ct):
            continue
        rango = _rango_filas(ws, fname, hoja)
        if not rango:
            continue
        for fila in range(rango[0], rango[1] + 1):
            cel = ws['{}{}'.format(col, fila)]
            v = cel.value
            if not isinstance(v, str) or not v.strip():
                continue
            clave = v.strip().lower()
            if clave in canon:
                continue                              # ya es canónica
            nueva = ALIAS_CATEGORIA.get(clave)
            if nueva and nueva != v:
                cel.value = nueva
                cambiadas += 1
                informe.append('{}:{}!{}{}: «{}» → «{}» (§1.1)'
                               .format(fname, hoja, col, fila, v, nueva))
    return cambiadas


# ==========================================================================
# §1.7 — formatos
# ==========================================================================
def _formato_de_cabecera(texto):
    if not isinstance(texto, str):
        return None
    t = texto.lower()
    for fmt, claves in FORMATO_POR_CABECERA:
        for k in claves:
            if k in t:
                return fmt
    return None


def aplicar_formatos(wb, fname, informe):
    """§1.7 — moneda, fecha, porcentaje y entero, deducidos de la cabecera.

    El `0` explícito de la columna «Días…» del 06 no es cosmética: sin él Excel
    hereda el formato de la celda de fecha que tiene al lado y `=J5-TODAY()`
    imprime «14/01/1900» en vez de «14».
    """
    n = 0
    for hoja, hdr, r0, r1_hoy, r1_v2 in BLOQUES.get(fname, []):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        rango = _rango_filas(ws, fname, hoja)
        if not rango:
            continue
        r0, r1 = rango
        for c in range(1, ws.max_column + 1):
            fmt = _formato_de_cabecera(ws.cell(row=hdr, column=c).value)
            if not fmt:
                continue
            for fila in range(r0, r1 + 1):
                cel = ws.cell(row=fila, column=c)
                if cel.number_format != fmt:
                    cel.number_format = fmt
                    n += 1
    for hoja, ref, fmt, cc, ct in RANGOS_FORMATO.get(fname, []):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        if not _centinela(ws, cc, ct):
            continue
        for fila in ws[ref]:
            for cel in (fila if isinstance(fila, tuple) else (fila,)):
                if cel.number_format != fmt:
                    cel.number_format = fmt
                    n += 1
    if n:
        informe.append('{}: {} celdas reformateadas (§1.7)'.format(fname, n))
    return n


# ==========================================================================
# Verde de edición: columnas SIN fórmula dentro del bloque de datos
# ==========================================================================
def aplicar_verde(wb, fname, informe):
    """Convención de familia: verde `E8F5E9` = lo escribe el cliente.

    Regla: dentro del bloque de datos, una columna es editable si tiene
    cabecera y NINGUNA de sus celdas contiene una fórmula. Las cabeceras de
    `NO_VERDE_CABECERAS` quedan fuera y las hojas de `SIN_VERDE_AUTO` las marcan
    los grupos a mano.

    La columna A lleva regla propia: es verde si su cabecera no es «#» y NO
    está pre-rellena entera. Así `04!'Control Recepción'!A` (Fecha, vacía) y
    `05!'Registro Diario Mermas'!A` (Fecha) SÍ son editables —lo son de verdad,
    el cliente escribe ahí— y en cambio `07!'Coste por Categoría'!A4:A13` y
    `05!'Análisis por Categoría'!A4:A10`, que son la taxonomía impresa, no.

    Y quita el verde de una columna que haya pasado a CALCULADA: si un grupo
    escribe la fórmula del «% del Total» en una columna que el motor pintó de
    verde en la pasada anterior, dejarla verde le diría al cliente que puede
    sobrescribir una fórmula. Se le devuelve el relleno de banda de su fila
    (el de la columna A), que es de donde salía.
    """
    n, limpiadas = 0, 0
    for hoja, hdr, r0d, r1h, r1v in BLOQUES.get(fname, []):
        if hoja not in wb.sheetnames or hoja in SIN_VERDE_AUTO:
            continue
        ws = wb[hoja]
        rango = _rango_filas(ws, fname, hoja)
        if not rango:
            continue
        r0, r1 = rango
        for c in range(1, ws.max_column + 1):
            cab = ws.cell(row=hdr, column=c).value
            if not isinstance(cab, str) or not cab.strip():
                continue
            if cab.strip().lower() in NO_VERDE_CABECERAS or cab.strip() == '#':
                continue
            celdas = [ws.cell(row=f, column=c) for f in range(r0, r1 + 1)]
            calculada = any(isinstance(x.value, str) and x.value.startswith('=')
                            for x in celdas)
            if c == 1 and all(x.value is not None for x in celdas):
                calculada = True               # columna de etiquetas impresa
            for cel in celdas:
                if calculada:
                    if es_verde(cel):
                        cel.fill = copy.copy(ws.cell(row=cel.row,
                                                     column=1).fill)
                        limpiadas += 1
                elif not es_verde(cel):
                    cel.fill = PatternFill('solid', fgColor=VERDE)
                    n += 1
    if n or limpiadas:
        informe.append('{}: {} celdas marcadas como editables (verde {}), '
                       '{} desmarcadas por pasar a calculadas'
                       .format(fname, n, VERDE, limpiadas))
    return n


# ==========================================================================
# §1.5 — protección sin contraseña
# ==========================================================================
def proteger(ws, informe):
    """Protección SIN contraseña: se desbloquean SÓLO las celdas verdes.

    Si la hoja no tiene ni una celda verde NO se protege: dejarla bloqueada
    entera es peor que dejarla abierta — el cliente no podría escribir nada y
    creería que el fichero está roto.
    """
    verdes = 0
    for row in ws.iter_rows():
        for c in row:
            if es_verde(c):
                c.protection = Protection(locked=False)
                verdes += 1
            else:
                c.protection = Protection(locked=True)
    if not verdes and ws.title not in PROTEGER_SIN_VERDE:
        ws.protection.sheet = False
        return 0
    ws.protection.sheet = True
    # SIN contraseña. Ojo: NO tocar `password`. `= None` revienta openpyxl y
    # `= ''` escribe el hash de la cadena vacía → Excel pediría contraseña al
    # desproteger, justo lo contrario de lo que dicen las Instrucciones.
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.autoFilter = False
    ws.protection.sort = False
    informe.append('{}: protegida sin contraseña ({} celdas verdes editables)'
                   .format(ws.title, verdes))
    return verdes


# ==========================================================================
# §1.10 — bio anclada, línea de versión y nota de desprotección
# ==========================================================================
def bio_y_version(wb, informe):
    """§1.10 — deja SIEMPRE tres líneas seguidas al pie de `Instrucciones`:
    nota de desprotección, bio anclada y línea de versión.

    Es una INSERCIÓN: los 9 ficheros terminan hoy en «Versión 1.1 · agosto
    2026 · …» y no llevan bio (por eso `postprocess-transversal.py`, que sólo
    SUSTITUYE patrones conocidos, nunca se la puso).

    Idempotencia: el ancla es la nota de desprotección — si ya está, el bloque
    se reescribe en su sitio. Anclar en la línea de versión desplazaría el
    bloque dos filas en cada pasada.
    """
    if 'Instrucciones' not in wb.sheetnames:
        return 0
    ws = wb['Instrucciones']
    col = 1
    fila_nota = fila_bio = fila_ver = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str):
            continue
        if v == NOTA_DESPROTEGER:
            fila_nota = r
        elif v == BIO:
            fila_bio = r
        elif RX_VERSION.match(v):
            fila_ver = r

    if fila_nota:
        inicio = fila_nota
    elif fila_bio:
        inicio = fila_bio - 1
    elif fila_ver:
        inicio = fila_ver
    else:
        inicio = ws.max_row + 2

    estilo = None
    if fila_ver:
        estilo = ws.cell(row=fila_ver, column=col)._style

    for i, texto in enumerate((NOTA_DESPROTEGER, BIO, VERSION_LINE)):
        cel = ws.cell(row=inicio + i, column=col)
        cel.value = texto
        if estilo is not None:
            cel._style = copy.copy(estilo)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
    # una línea de versión perdida en otra fila sería un duplicado
    for r in range(1, ws.max_row + 1):
        if inicio <= r <= inicio + 2:
            continue
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and (RX_VERSION.match(v) or v == BIO
                                   or v == NOTA_DESPROTEGER):
            ws.cell(row=r, column=col).value = None
    informe.append('Instrucciones!A{}:A{}: nota de desprotección + bio '
                   'anclada + versión {}'.format(inicio, inicio + 2, VERSION))
    return 3


# ==========================================================================
# §1.10 — metadata OOXML
# ==========================================================================
def set_metadata(wb, fname, informe):
    """Propiedades que ve el cliente en Archivo → Información. `subject` pasa a
    `… · v2.0`; el resto sigue la convención de `postprocess-transversal.py`
    (`creator='AI Chef Pro'`, no 'openpyxl')."""
    p = wb.properties
    titulo = p.title if isinstance(p.title, str) else ''
    if not titulo.strip():
        base = re.sub(r'^(BONUS-)?\d+[a-z]?-', '', fname[:-5])
        titulo = base.replace('-', ' ').capitalize()
    if not titulo.endswith(CORTO):
        titulo = '{} · {}'.format(titulo.split(' · ')[0], CORTO)
    quiero = dict(
        creator='AI Chef Pro',
        lastModifiedBy='AI Chef Pro',
        title=titulo,
        subject='{} · v{}'.format(CORTO, VERSION),
        keywords='{}, AI Chef Pro'.format(PID.replace('-', ' ')),
        description='aichef.pro/{}'.format(PID),
        category='AI Chef Pro · Productos digitales',
    )
    n = 0
    for campo, valor in quiero.items():
        if getattr(p, campo) != valor:
            setattr(p, campo, valor)
            n += 1
    if n:
        informe.append('{}: {} propiedades OOXML actualizadas (subject → v{})'
                       .format(fname, n, VERSION))
    return n


# ==========================================================================
# Contadores y gates auxiliares
# ==========================================================================
RX_ENTRECOMILLADO = re.compile(r"['‘’\"“”]([^'‘’\"“”]{3,40})['‘’\"“”]")


RX_ENUMERADA = re.compile(r'^\s*\d+\.\s*(.+?)\s+[—–-]\s')


def _cabeceras(wb):
    """Todos los textos que son ENCABEZADO de columna en alguna hoja. Sirven
    para no confundir «'Producto'» o «'Estado'» —que son columnas— con una
    pestaña inexistente."""
    fuera = set()
    for ws in wb.worksheets:
        for fila in range(1, min(9, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=fila, column=c).value
                if isinstance(v, str) and 1 <= len(v) <= 40:
                    fuera.add(v.strip().lower())
    return fuera


def pestanas_citadas(wb):
    """§1.6 — gate: toda pestaña citada en `Instrucciones` tiene que existir en
    `wb.sheetnames`.

    Hoy falla en tres formas distintas, y por eso el gate mira tres cosas:

    * ENTRECOMILLADA con la palabra «pestaña» al lado — `03!A9` promete
      «Imprimible», que no existe, y `01!A20` dice «Resumen» donde la hoja se
      llama «Resumen Dashboard».
    * ENUMERADA en un bloque «PESTAÑAS:» con el patrón `N. Nombre — …` —
      `02!A7`/`A9` dicen «Directorio» y «Evaluación» (son «Directorio
      Proveedores» y «Evaluación Proveedores») y `07!A8` cita «Evolución
      Mensual», una hoja que sencillamente no está.
    * CITADA entre comillas sin la palabra pestaña — `06!A11` manda usar el
      «Mapa de Almacén» y la hoja se llama «Mapa Almacén».

    Buscar sólo lo entrecomillado con la palabra «pestaña» delante encuentra 3
    de las 7. Devuelve [(fila, nombre_citado)].
    """
    if 'Instrucciones' not in wb.sheetnames:
        return []
    ws = wb['Instrucciones']
    reales = set(h.strip().lower() for h in wb.sheetnames)
    cabeceras = _cabeceras(wb)
    fuera = []
    en_bloque = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        if re.search(r'PESTA[ÑN]AS', v, re.I):
            en_bloque = True
            continue
        if en_bloque:
            m = RX_ENUMERADA.match(v)
            if m:
                nombre = m.group(1).strip()
                if nombre.lower() not in reales:
                    fuera.append((r, nombre))
                continue
            if v.strip():
                en_bloque = False
        for m in RX_ENTRECOMILLADO.finditer(v):
            nombre = m.group(1).strip()
            if not nombre or nombre.lower() in reales:
                continue
            # «'Producto'», «'Estado'», «'A Pedir'»… son COLUMNAS, no pestañas
            if nombre.lower() in cabeceras:
                continue
            fuera.append((r, nombre))
    return fuera


def hojas_esqueleto(wb, fname):
    """§1.4 — contador: hojas cuyo bloque de datos está COMPLETAMENTE en blanco
    (ni un dato, ni un ejemplo). Son las 12 que la SPEC manda sembrar."""
    fuera = []
    for hoja, hdr, r0, r1_hoy, r1_v2 in BLOQUES.get(fname, []):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        rango = _rango_filas(ws, fname, hoja)
        if not rango:
            fuera.append(hoja)     # ni una fila de datos: cabecera y nada más
            continue
        lleno = False
        for fila in range(rango[0], rango[1] + 1):
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=fila, column=c).value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith('='):
                    continue
                lleno = True
                break
            if lleno:
                break
        if not lleno:
            fuera.append(hoja)
    return fuera


def contar(wb):
    """Censo interno de un libro: lo que main.py vuelca al informe."""
    r = dict(hojas=len(wb.sheetnames), formulas=0, verdes=0,
             verdes_con_dato=0, dv=0, cf=0, protegidas=0, celdas=0)
    for ws in wb.worksheets:
        r['dv'] += len(ws.data_validations.dataValidation)
        r['cf'] += sum(len(cf.rules) for cf in ws.conditional_formatting)
        if ws.protection.sheet:
            r['protegidas'] += 1
        for row in ws.iter_rows():
            for c in row:
                # el verde se cuenta AUNQUE la celda esté vacía: una casilla
                # editable en blanco es precisamente lo que se le ofrece al
                # cliente. Contar sólo las que tienen dato daba 250 donde el
                # fichero tiene 513.
                if es_verde(c):
                    r['verdes'] += 1
                    if c.value is not None:
                        r['verdes_con_dato'] += 1
                if c.value is None:
                    continue
                r['celdas'] += 1
                if isinstance(c.value, str) and c.value.startswith('='):
                    r['formulas'] += 1
    return r


# ==========================================================================
# Entrada del pipeline
# ==========================================================================
def aplicar(wb, fname, informe):
    """PRE-pase del motor (§1), antes de que el grupo escriba su contenido.

    Aquí sólo va lo que NO depende de que los grupos hayan movido columnas:
    la normalización de la taxonomía y la tabla de IVA del 03. Todo lo demás
    (DV, formato condicional, formatos, verde, protección, bio, metadata) vive
    en `cerrar()`, que corre DESPUÉS del grupo.
    """
    normalizar_categorias(wb, fname, informe)
    if fname.startswith('03-'):
        tabla_iva(wb, informe)
    return informe


def cerrar(wb, fname, informe, proteger_hojas=True):
    """POST-pase del motor: lo que depende del layout FINAL.

    Devuelve el dict de gates del fichero (pendientes de grupo, pestañas
    citadas inexistentes, hojas esqueleto, contadores).
    """
    pendientes = []

    # limpieza previa: la cadena vacía es un defecto del censo (`empty_str`)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    c.value = None
        _limpiar_dv(ws)

    aplicar_formatos(wb, fname, informe)
    aplicar_verde(wb, fname, informe)
    n_dv = aplicar_dv(wb, fname, informe, pendientes)
    n_cf = aplicar_cf(wb, fname, informe, pendientes)

    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            print_setup(ws, None, landscape=False)
            continue
        b = _bloque(fname, ws.title)
        print_setup(ws, b[0] if b else None,
                    landscape=ws.max_column >= 6)
        ultima = get_column_letter(max(1, ws.max_column))
        ws.print_area = 'A1:{}{}'.format(ultima, max(1, ws.max_row))
        # Ronda 2 · RT-13 — `Listas` ya NO se exceptúa: es la tabla de IVA de
        # la que vive el VLOOKUP de `Pedido Actual!H`, que sí está protegida.
        # Sus celdas de tipo y de producto son verdes, así que siguen siendo
        # editables con la hoja protegida.
        if proteger_hojas:
            proteger(ws, informe)

    bio_y_version(wb, informe)
    set_metadata(wb, fname, informe)

    citadas = pestanas_citadas(wb)
    esqueleto = hojas_esqueleto(wb, fname)
    return {
        'fichero': fname,
        'dv_aplicadas': n_dv,
        'cf_aplicadas': n_cf,
        'pendientes_de_grupo': pendientes,
        'pestanas_citadas_inexistentes': ['Instrucciones!A{}: «{}»'
                                          .format(f, n) for f, n in citadas],
        'hojas_esqueleto': esqueleto,
        'contadores': contar(wb),
    }
