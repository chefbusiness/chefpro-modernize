#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
grupo_a.py — §2 de `kit-inventario-v2-SPEC.md`: inventario y proveedores.

Ficheros: `01-inventario-stock-diario.xlsx`, `02-fichas-proveedores.xlsx` y
`BONUS-08-inventario-rapido-mensual.xlsx`.

Qué arregla (ids del R1, `auditorias/kit-inventario-R1.json`):

  01  DOM-01/DOM-02/TEC-01/COM-02  las 50 filas precargadas tenían categoría y
      unidad asignadas por ROTACIÓN CÍCLICA (la pechuga de pollo era
      «Pescados» y el solomillo se medía en litros). Se reescriben producto a
      producto con su categoría canónica, su unidad REAL de compra y un precio
      HORECA orientativo.
      DOM-14/TEC-25  los 50 par levels eran 5/15 sin excepción: ahora van por
      producto y formato.
      DOM-09/TEC-03/COM-14  la valoración de stock no existía (columna «Valor
      (€)» vacía, sin precio con el que calcularla y `Resumen Dashboard` con 0
      fórmulas). Se inserta `J Precio/ud (€)` y se construye el dashboard.
      DOM-33/TEC-21  «A Pedir» decía 0 en toda la banda ámbar.
      COM-31  «Coca-Cola» → «Refresco de cola».
  02  DOM-17  la ficha de proveedor no recogía CIF/NIF, RGSEAA ni homologación.
      DOM-15/DOM-16/TEC-07/COM-18  la comparativa daba 0 en las 10 filas
      (`MIN` sobre rango vacío), no decía QUIÉN era el más barato y comparaba
      la garrafa de 5 L contra el litro.
      TEC-16/COM-29  «TOTAL» que era una media, guarda de un solo criterio y
      columna «Nota» vacía.
  B08 DOM-32/TEC-24/COM-28  el inventario mensual no calculaba el CONSUMO —que
      es para lo que se hace— y la «Variación» presentaba el stock entero como
      crecimiento la primera vez que se usa la plantilla.

Contrato con `main.py`: `FICHEROS`, `pre(wb, fname, cambios)`,
`post(wb, fname, cambios, registro)` y `demos(carpeta, origen)`. Las columnas
se insertan en `pre()` (antes de que el motor fije rangos); las filas se AÑADEN
en `post()`. Toda fórmula pasa por `motor._reg` para que `main.py` verifique su
valor cacheado.

Python 3.7 / openpyxl 3.1.3: sin walrus ni f-strings de depuración.
"""
import contextlib
import copy
import datetime
import os

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

import motor

FICHEROS = [
    '01-inventario-stock-diario.xlsx',
    '02-fichas-proveedores.xlsx',
    'BONUS-08-inventario-rapido-mensual.xlsx',
]

#: Marca de las validaciones que pone ESTE grupo. NO puede empezar por
#: `motor.MARCA_DV` ('kitinv-v2'): `motor._limpiar_dv()` corre dentro de
#: `cerrar()` —después de `post()`— y borraría todo lo que lleve esa marca.
MARCA_A = 'kitinv-a'

NOTA_EJEMPLO = ('Datos de ejemplo; precios orientativos SIN IVA, edítalos con '
                'los tuyos. Categoría y Unidad tienen desplegable.')


# ==========================================================================
# §1.7 — una regla de formato más para el motor
# ==========================================================================
def _extender_formatos():
    """«Compras del mes (uds)» contiene «compras», que en
    `motor.FORMATO_POR_CABECERA` es palabra de MONEDA: sin esto el conteo de
    unidades del BONUS-08 se imprimiría como «40,00 €». La marca `(uds)` es
    explícita y no colisiona con «Precio/ud (€)», que no la lleva."""
    regla = (motor.FMT_CANT, ('(uds)',))
    if regla not in motor.FORMATO_POR_CABECERA:
        motor.FORMATO_POR_CABECERA.insert(0, regla)


_extender_formatos()


# ==========================================================================
# Datos precargados del 01 — categoría CANÓNICA, unidad REAL de compra,
# par level / par max por producto y precio HORECA orientativo (sin IVA).
# ==========================================================================
#: (producto, categoría, unidad, par level, par max, precio/ud €)
#:
#: Criterio de categoría, producto a producto (no hay ciclos):
#:   * aceites, vinagres, sal, especias, arroz, pasta, harina, azúcar y café
#:     son ECONOMATO SECO → «Secos/Granos».
#:   * los huevos van con «Lácteos» (es donde los pone el IVA del 4 % y donde
#:     están en el economato refrigerado).
#:   * el hielo va en «Congelados» (cámara de congelación, IVA 10 %).
#:   * papel, film, guantes, químicos y desechables → «Limpieza», que en la
#:     taxonomía del kit incluye menaje y desechables (§1.1).
#:   * el material de sala que no es ni alimento ni limpieza (GN, etiquetas,
#:     rollos de ticket) → «Otros».
PRODUCTOS = [
    ('Cocina', [
        ('Pechuga de pollo',                'Cárnicos',               'kg',      8,  20,  6.50),
        ('Solomillo de ternera',            'Cárnicos',               'kg',      3,   8, 32.00),
        ('Salmón fresco',                   'Pescados',               'kg',      4,  10, 14.00),
        ('Gambas',                          'Pescados',               'kg',      3,   8, 18.00),
        ('Tomate',                          'Verduras/Frutas',        'kg',     10,  25,  1.95),
        ('Cebolla',                         'Verduras/Frutas',        'kg',     10,  25,  1.10),
        ('Lechuga',                         'Verduras/Frutas',        'ud',     12,  30,  0.95),
        ('Patata',                          'Verduras/Frutas',        'kg',     25,  60,  0.85),
        ('Aceite de oliva virgen extra',    'Secos/Granos',           'L',      10,  25,  5.80),
        ('Sal marina',                      'Secos/Granos',           'kg',      3,  10,  0.60),
        ('Pimienta negra molida',           'Secos/Granos',           'kg',      1,   3, 12.00),
        ('Arroz redondo',                   'Secos/Granos',           'kg',     10,  25,  1.45),
        ('Pasta seca',                      'Secos/Granos',           'kg',      8,  20,  1.35),
        ('Harina de trigo',                 'Secos/Granos',           'kg',     10,  25,  0.75),
        # RD-04 · los huevos NO son lácteos: el mapa de almacén manda los
        # lácteos a la cámara de elaborados (0-4 °C) y refrigerar el huevo
        # antes de la venta es lo que prohíbe el Reg. (CE) 589/2008 —la
        # condensación en la cáscara facilita la entrada de Salmonella—.
        # Ninguna de las 10 categorías canónicas les sirve, así que van a
        # «Otros» y el mapa del 06 estrena su zona de ambiente estable.
        ('Huevos M',                        'Otros',                  'docena', 15,  40,  2.20),
        ('Nata 35 % M.G.',                  'Lácteos',                'L',       6,  15,  3.10),
        ('Mantequilla',                     'Lácteos',                'kg',      3,   8,  8.90),
        ('Queso parmesano',                 'Lácteos',                'kg',      2,   5, 16.50),
        ('Limón',                           'Verduras/Frutas',        'kg',      5,  12,  1.80),
        ('Ajo',                             'Verduras/Frutas',        'kg',      2,   6,  4.20),
    ]),
    ('Barra', [
        ('Café en grano',                   'Secos/Granos',           'kg',      6,  15, 14.50),
        ('Leche entera',                    'Lácteos',                'L',      24,  60,  0.85),
        ('Bebida de avena',                 'Bebidas No Alcohólicas', 'L',       6,  18,  1.60),
        ('Zumo de naranja',                 'Bebidas No Alcohólicas', 'L',       8,  20,  1.95),
        ('Refresco de cola',                'Bebidas No Alcohólicas', 'ud',     48, 120,  0.55),
        ('Cerveza de grifo (barril 30 L)',  'Bebidas Alcohólicas',    'barril',  2,   6, 55.00),
        ('Agua mineral',                    'Bebidas No Alcohólicas', 'ud',     60, 150,  0.28),
        ('Tónica',                          'Bebidas No Alcohólicas', 'ud',     24,  72,  0.58),
        ('Vino tinto (botella 75 cl)',      'Bebidas Alcohólicas',    'ud',     12,  36,  4.80),
        ('Vino blanco (botella 75 cl)',     'Bebidas Alcohólicas',    'ud',     12,  36,  4.50),
        ('Hielo en cubitos',                'Congelados',             'saco',    6,  20,  1.80),
        ('Azúcar blanquilla',               'Secos/Granos',           'kg',      8,  20,  0.95),
        ('Servilletas de papel (barra)',    'Limpieza',               'caja',    4,  10, 12.50),
        ('Pajitas de papel',                'Limpieza',               'caja',    2,   6,  9.80),
        ('Vasos desechables',               'Limpieza',               'paquete', 6,  18,  3.40),
    ]),
    ('Almacén', [
        ('Papel de cocina',                 'Limpieza',               'rollo',  12,  30,  1.15),
        ('Film transparente',               'Limpieza',               'rollo',   4,  10,  4.60),
        ('Papel de aluminio',               'Limpieza',               'rollo',   4,  10,  5.20),
        ('Bolsas de basura',                'Limpieza',               'paquete', 6,  15,  3.90),
        ('Guantes de nitrilo',              'Limpieza',               'caja',    8,  20,  6.50),
        ('Detergente de lavavajillas',      'Limpieza',               'L',      10,  25,  2.40),
        ('Desengrasante de cocina',         'Limpieza',               'L',       6,  15,  3.10),
        ('Lejía alimentaria',               'Limpieza',               'L',       6,  15,  1.25),
        ('Cubetas GN 1/1',                  'Otros',                  'ud',      6,  15, 18.00),
        ('Etiquetas FIFO',                  'Otros',                  'rollo',   3,   8,  7.50),
        ('Rollos de ticket',                'Otros',                  'paquete', 5,  12,  4.20),
        # RD-25 · la misma referencia dada de alta en dos zonas: es stock
        # POR UBICACIÓN, y ahora el nombre lo dice y las Instrucciones lo
        # explican (el pedido se consolida en la plantilla 03).
        ('Servilletas de papel (reserva de almacén)',
                                            'Limpieza',               'caja',    3,   8, 12.50),
        ('Aceite de girasol (garrafa 5 L)', 'Secos/Granos',           'L',      15,  40,  1.85),
        ('Vinagre de vino',                 'Secos/Granos',           'L',       6,  15,  1.40),
        ('Legumbres secas',                 'Secos/Granos',           'kg',     10,  25,  1.60),
    ]),
]

#: (hoja, última fila de datos en v1.1, última fila en v2.0, fila del pie)
ZONAS_01 = [('Cocina', 24, 44, 27), ('Barra', 19, 34, 22),
            ('Almacén', 19, 34, 22)]


def _gate_coherencia_consumos():
    """RD-02/RD-21 · GATE de importación: el par level de cada producto del 01
    tiene que ser EXACTAMENTE `consumo diario × (lead time + cobertura)` y el
    par max su «stock máximo», porque de esa misma tabla salen el punto de
    pedido del BONUS-09 y el gasto mensual del Top 20 del 07.

    Va aquí, a nivel de módulo, para que el descuadre reviente al IMPORTAR y no
    tres pantallas más allá dentro de un .xlsx que nadie va a volver a abrir:
    el kit daba tres puntos de pedido distintos para el mismo producto
    precisamente porque nada lo comprobaba.
    """
    catalogo = {}
    for _hoja, filas in PRODUCTOS:
        for nombre, _cat, _ud, par, par_max, precio in filas:
            catalogo[nombre] = (par, par_max, precio)
    problemas = []
    for nombre, datos in motor.CONSUMO_EJEMPLO.items():
        consumo, lead, cobertura, _vida, stock_max = datos
        if nombre not in catalogo:
            problemas.append('{}: está en motor.CONSUMO_EJEMPLO y NO en las 50 '
                             'filas del 01'.format(nombre))
            continue
        par, par_max, _precio = catalogo[nombre]
        derivado = consumo * (lead + cobertura)
        if abs(derivado - par) > 1e-9:
            problemas.append('{}: par level del 01 = {} pero consumo × (lead + '
                             'cobertura) = {}'.format(nombre, par, derivado))
        if stock_max != par_max:
            problemas.append('{}: par MAX del 01 = {} y stock máximo de '
                             'CONSUMO_EJEMPLO = {}'.format(nombre, par_max,
                                                           stock_max))
    if problemas:
        raise AssertionError(
            'kit-inventario · los datos de ejemplo del 01, del 07 y del '
            'BONUS-09 han dejado de cuadrar (RD-02/RD-21):\n  '
            + '\n  '.join(problemas))


_gate_coherencia_consumos()


# ==========================================================================
# Datos precargados del 02 — una ficha de ejemplo por TIPO de proveedor
# ==========================================================================
#: Cabecera final del `Directorio Proveedores` (A3:S3). Las 8 columnas nuevas
#: son DOM-17: sin CIF/NIF no se da de alta al proveedor en contabilidad y sin
#: Nº RGSEAA no se acredita el prerrequisito de homologación de proveedores del
#: plan APPCC. «Días Entrega» (vieja H) pasa a llamarse «Día de reparto», que
#: es lo que en realidad se anota.
CAB_DIRECTORIO = [
    '#', 'Proveedor', 'Categoría', 'CIF/NIF', 'Nº RGSEAA', 'Contacto',
    'Teléfono', 'Email', 'Contacto de incidencias', 'Dirección',
    'Día de pedido', 'Día de reparto', 'Plazo de entrega (días)',
    'Pedido Mínimo (€)', 'Condiciones de pago', 'Homologado (S/N)',
    'Fecha de homologación', 'Fecha de última revisión', 'Notas',
]

ANCHOS_DIRECTORIO = [5, 26, 20, 13, 20, 18, 14, 26, 26, 32, 20, 14, 12, 14,
                     20, 14, 14, 14, 40]

#: Posiciones (en coordenadas ORIGINALES) donde hay que insertar columna, de
#: mayor a menor para que los índices sigan siendo válidos al aplicarlas en
#: cadena. Original: A# B Prov C Cat D Contacto E Tel F Email G Dir
#: H DíasEntrega I PedidoMín J FormaPago K Notas.
INSERTS_DIRECTORIO = [11, 11, 11, 9, 8, 7, 4, 4]

#: Fichas de ejemplo (una por familia de compra). Empresas ficticias: el
#: dominio `ejemplo-*.es` deja claro que son de muestra y el Nº RGSEAA enseña
#: el FORMATO (clave.número/provincia) — el real lo da el proveedor.
PROVEEDORES = [
    ('Cárnicas del Norte', 'Cárnicos', 'B31245678', '10.05412/NA',
     'Javier Beltrán', '948 21 34 55', 'pedidos@ejemplo-carnicas.es',
     'incidencias@ejemplo-carnicas.es · 948 21 34 60',
     'Pol. Ind. Landaben, nave 12 · 31012 Pamplona',
     'Lun y jue antes de 12:00', 'Mar y vie', 1, 150,
     '30 días fecha factura', 'S',
     datetime.date(2026, 2, 10), datetime.date(2026, 7, 15),
     'Exigir albarán con nº de lote y temperatura de llegada.'),
    ('Pescados Ría Fresca', 'Pescados', 'B36998877', '12.00987/PO',
     'Marta Souto', '986 44 12 09', 'pedidos@ejemplo-pescados.es',
     'incidencias@ejemplo-pescados.es · 986 44 12 15',
     'Mercado Central, puesto 8 · 36202 Vigo',
     'Diario antes de 17:00', 'Mar a sáb', 1, 120,
     '15 días fecha factura', 'S',
     datetime.date(2026, 3, 5), datetime.date(2026, 7, 20),
     'Llega en hielo fundente; rechazar por encima de 2 °C.'),
    ('Frutas y Verduras La Huerta', 'Verduras/Frutas', 'B46112233',
     '21.024567/V', 'Ana Ferrer', '963 55 71 20',
     'pedidos@ejemplo-huerta.es', 'incidencias@ejemplo-huerta.es',
     'Mercavalencia, nave 4 · 46009 Valencia',
     'Lun, mié y vie antes de 18:00', 'Mar, jue y sáb', 1, 80,
     'Contado', 'S',
     datetime.date(2026, 1, 12), datetime.date(2026, 7, 12),
     'Producto de temporada: confirmar precio cada semana.'),
    ('Distribuciones Economato Sur', 'Secos/Granos', 'B41778899',
     # RD-26 · la primera cifra del RGSEAA identifica la ACTIVIDAD: 21 son
     # vegetales, 30 bebidas alcohólicas y 40 almacenamiento y distribución
     # polivalente. Tres claves 21 para tres sectores distintos enseñaban un
     # número que no puede existir tal cual, justo en el dato con el que se
     # acredita la homologación ante una inspección.
     '40.031204/SE', 'Rafael Ortega', '954 33 90 41',
     'pedidos@ejemplo-economato.es', 'incidencias@ejemplo-economato.es',
     'Pol. Ind. La Red, nave 27 · 41500 Alcalá de Guadaíra',
     'Mar antes de 14:00', 'Jue', 2, 250,
     '45 días fecha factura', 'S',
     datetime.date(2026, 2, 20), datetime.date(2026, 7, 18),
     'Rappel del 2 % a partir de 2.000 € al mes.'),
    ('Bebidas y Distribución Levante', 'Bebidas Alcohólicas', 'B03445566',
     '30.040918/A', 'Luis Cano', '965 12 78 30',
     'pedidos@ejemplo-bebidas.es', 'incidencias@ejemplo-bebidas.es',
     'Ctra. N-340, km 82 · 03008 Alicante',
     'Lun antes de 12:00', 'Mié', 2, 200,
     '30 días fecha factura', 'S',
     datetime.date(2026, 1, 28), datetime.date(2026, 7, 10),
     'Cesión de barriles y CO2: revisar el depósito de envases.'),
    # RD-19 · este proveedor saca un 2,2 en la evaluación («D — Sustituir») y
    # el directorio lo mantenía homologado con la revisión MÁS RECIENTE de los
    # seis. Enseñar como ejemplo un proveedor suspendido y homologado a la vez
    # enseña justo lo contrario de lo que el kit dice hacer.
    ('Higiene Profesional HORECA', 'Limpieza', 'B28556677',
     'No aplica (producto no alimentario)', 'Silvia Gómez', '916 40 22 18',
     'pedidos@ejemplo-higiene.es', 'incidencias@ejemplo-higiene.es',
     'Av. de la Industria 15 · 28108 Alcobendas',
     'Mié antes de 16:00', 'Vie', 2, 100,
     '30 días fecha factura', 'Pendiente',
     datetime.date(2026, 2, 3), datetime.date(2026, 8, 3),
     'Evaluación D (2,2/5) en la última revisión: homologación EN SUSPENSO '
     'hasta que mejore. Buscar alternativa antes de la próxima revisión.'),
]

CAB_CONDICIONES = [
    '#', 'Proveedor', 'Plazo de pago', 'Pedido Mínimo (€)', 'Rappel (%)',
    'Transporte', 'Horario de entrega', 'Días de pedido', 'Días de reparto',
    'Notas',
]

#: (plazo de pago, pedido mínimo, rappel, transporte, horario, días de pedido,
#:  días de reparto, notas) — mismo orden que `PROVEEDORES`.
CONDICIONES = [
    ('30 días fecha factura', 150, 0.02, 'Incluido desde 150 €',
     '07:00 - 09:00', 'Lun y jue', 'Mar y vie',
     'Portes de 12 € por debajo del mínimo.'),
    ('15 días fecha factura', 120, 0.00, 'Incluido', '06:30 - 08:00',
     'Diario', 'Mar a sáb', 'Subasta: el precio se cierra la víspera.'),
    ('Contado', 80, 0.00, 'Incluido desde 80 €', '06:00 - 08:00',
     'Lun, mié y vie', 'Mar, jue y sáb', 'Sin pedido mínimo los sábados.'),
    ('45 días fecha factura', 250, 0.02, '18 € por debajo del mínimo',
     '09:00 - 13:00', 'Mar', 'Jue', 'Rappel liquidado por trimestres.'),
    ('30 días fecha factura', 200, 0.03, 'Incluido desde 200 €',
     '10:00 - 13:00', 'Lun', 'Mié', 'Depósito de barriles y botelleros.'),
    ('30 días fecha factura', 100, 0.00, 'Incluido desde 100 €',
     '11:00 - 14:00', 'Mié', 'Vie', 'Fichas de seguridad en cada entrega.'),
]

#: Puntuaciones de ejemplo (calidad, precio, puntualidad, servicio,
#: flexibilidad). Cubren a propósito las cuatro notas A/B/C/D.
EVALUACIONES = [
    (5, 4, 5, 5, 5),
    (5, 3, 4, 4, 3),
    (4, 4, 3, 4, 4),
    (3, 5, 4, 3, 2),
    (4, 4, 5, 4, 4),
    (2, 3, 2, 2, 2),
]

#: Comparativa de precios: PRODUCTO, unidad, formato de venta y contenido de
#: cada línea. Es lo que arregla DOM-16 — sin normalizar, `MIN` premia a quien
#: cotiza la garrafa de 5 L frente a quien cotiza el litro.
#:
#: Ronda 2 · RC-20 — los nombres son EXACTAMENTE los del 01: la hoja arrastraba
#: el sufijo redundante «(kg)» (el formato de venta ya va en su columna) y
#: renombraba el café, así que la única hoja del kit que rompía la promesa de
#: vocabulario común lo hacía sin necesidad.
#: (producto del 01, unidad, formato de venta, contenido, precio/ud de
#:  referencia del 01)
COMPARATIVA = [
    ('Pechuga de pollo',             'kg',     'Caja 5 kg',              5,  6.50),
    ('Solomillo de ternera',         'kg',     'Pieza al vacío ~2 kg',   2, 32.00),
    ('Salmón fresco',                'kg',     'Pieza entera ~4 kg',     4, 14.00),
    ('Aceite de oliva virgen extra', 'L',      'Garrafa 5 L',            5,  5.80),
    ('Tomate',                       'kg',     'Caja 10 kg',            10,  1.95),
    ('Lechuga',                      'ud',     'Caja 12 ud',            12,  0.95),
    ('Arroz redondo',                'kg',     'Saco 20 kg',            20,  1.45),
    ('Huevos M',                     'docena', 'Caja 12 docenas',       12,  2.20),
    ('Leche entera',                 'L',      'Caja 12 briks de 1 L',  12,  0.85),
    ('Queso parmesano',              'kg',     'Caja 4 kg',              4, 16.50),
]

#: Cotizaciones de ejemplo: fila del bloque (0 = fila 4) → 5 precios + fecha.
#:
#: Ronda 2 · RD-11 — el AOVE costaba 5,80 €/L en el 01 y en el 07 y 8,40 €/L
#: aquí: un 45 % de diferencia para la MISMA referencia dentro del mismo kit,
#: y justo en la hoja que dice venir a evitar ese error. Ahora el precio
#: ganador de cada línea es EXACTAMENTE `precio de referencia × contenido`,
#: así que el precio/ud que sale de la comparativa es el que el 01 valora y el
#: 07 analiza. Y el ganador es el proveedor de la familia del producto (la
#: columna F es Cárnicas, la H la frutería, la I el economato), no uno al azar.
COTIZACIONES = {
    0: ([32.50, 33.50, 34.20, None, 33.00], datetime.date(2026, 8, 3)),
    3: ([31.00, 30.50, 30.20, 29.00, 31.50], datetime.date(2026, 8, 3)),
    4: ([20.50, 21.00, 19.50, 20.20, 19.90], datetime.date(2026, 8, 10)),
}


# ==========================================================================
# Utilidades del grupo
# ==========================================================================
def _f(ws, coord, formula):
    """Escribe una fórmula y la registra para que `main.py` compruebe que
    quedó con valor cacheado."""
    ws[coord] = motor._reg(ws, coord, formula)
    return formula


def _limpiar_dv_a(ws):
    """Quita las DV de ESTE grupo antes de reescribirlas. Sin esto la 2.ª
    pasada acumularía una copia y la idempotencia daría «cambia dv»."""
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if not (getattr(dv, 'promptTitle', None) or '').startswith(MARCA_A)]


def _dv(ws, ref, valores, titulo, prompt):
    dv = DataValidation(
        type='list', formula1='"{}"'.format(','.join(valores)),
        allow_blank=True, showErrorMessage=True, errorStyle='stop',
        errorTitle=titulo, error='Elige un valor de la lista.',
        showInputMessage=True,
        promptTitle='{} · {}'.format(MARCA_A, titulo), prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(ref)
    return dv


def _cabecera(ws, fila, textos, col0=1):
    """Reescribe una fila de cabecera copiando el estilo de la primera celda
    que ya lo tenga (la cabecera oscura del kit)."""
    modelo = ws.cell(row=fila, column=col0)._style
    for i, txt in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i)
        cel.value = txt
        cel._style = copy.copy(modelo)


def _nota(ws, coord, texto):
    cel = ws[coord]
    cel.value = texto
    cel.font = Font(italic=True, size=9, color='666666')
    cel.alignment = Alignment(vertical='center')


def _instrucciones(ws, lineas):
    """Reescribe `Instrucciones` de arriba abajo.

    IDEMPOTENCIA: la última línea es `motor.VERSION_LINE`, que es donde
    `motor.bio_y_version()` —que corre después, en `cerrar()`— ancla su bloque
    de tres líneas. En la 2.ª pasada esta función vuelve a dejar el fichero
    como estaba antes de `cerrar()`, y `cerrar()` reescribe lo mismo. Lo que
    NO se puede hacer es APPEND al final: `bio_y_version` escribiría encima.
    """
    modelo = ws.cell(row=4, column=1)._style
    maximo = ws.max_row
    for i, txt in enumerate(lineas):
        cel = ws.cell(row=i + 1, column=1)
        if cel.value is None and txt is not None:
            cel._style = copy.copy(modelo)
        cel.value = txt
    for r in range(len(lineas) + 1, maximo + 1):
        ws.cell(row=r, column=1).value = None


def _ancho(ws, anchos, col0=1):
    for i, a in enumerate(anchos):
        ws.column_dimensions[
            motor.get_column_letter(col0 + i)].width = a


# ==========================================================================
# 01 — inventario de stock diario
# ==========================================================================
#: DOM-33/TEC-21 — «A Pedir» arranca ya en ÁMBAR (por debajo de 1,5 × par
#: level), no solo en rojo, y repone hasta el Par Max. `MAX(0;…)` evita que un
#: stock por encima del par max devuelva un negativo.
#: RT-07 — la fórmula de la v2.0 sólo guardaba el stock contado, no el par
#: level. En las 60 filas LIBRES que añade §1.9 —las que Instrucciones!A19
#: invita expresamente a usar— el par level está vacío, así que cualquier
#: cantidad que el cliente escribiera comparaba contra un vacío (=0) y salía
#: «🟢 OK», y «A Pedir» salía 0: el producto que da de alta el primer día es
#: precisamente el que nunca le iba a avisar. Con la guarda, la fila nueva se
#: queda en blanco hasta que tenga par level, que es lo honesto.
F_ESTADO = ('=IF(OR($G{f}="",$E{f}=""),"",IF($G{f}<$E{f},"🔴 PEDIR",'
            'IF($G{f}<$E{f}*1.5,"🟡 BAJO","🟢 OK")))')
F_APEDIR = ('=IF(OR($G{f}="",$E{f}="",$F{f}=""),"",'
            'IF($G{f}<$E{f}*1.5,MAX(0,$F{f}-$G{f}),0))')
#: DOM-09/TEC-03/COM-14 — valoración de stock: cantidad contada × precio.
F_VALOR = '=IFERROR(IF(OR($G{f}="",$J{f}=""),"",$G{f}*$J{f}),"")'


def _pre_01(wb, cambios):
    """Inserta `J Precio/ud (€)`: sin precio unitario la columna «Valor (€)»
    no puede existir (DOM-09). Valor→K, Proveedor→L, Notas→M."""
    for hoja, r_ult, r_v2, pie in ZONAS_01:
        ws = wb[hoja]
        if 'Precio' in str(ws['J4'].value or ''):
            continue                                   # ya insertada
        motor.insertar_columna(ws, 10)
        cambios.append('01:{}!J: columna «Precio/ud (€)» insertada '
                       '(Valor→K, Proveedor→L, Notas→M) — DOM-09'.format(hoja))


def _post_01(wb, cambios):
    for hoja, r_ult, r_v2, pie in ZONAS_01:
        ws = wb[hoja]
        filas = dict(PRODUCTOS)[hoja]
        _nota(ws, 'A3', NOTA_EJEMPLO)
        ws['J4'] = 'Precio/ud (€)'
        ws['J4']._style = copy.copy(ws['I4']._style)
        for i, dato in enumerate(filas):
            f = 5 + i
            producto, categoria, unidad, par, par_max, precio = dato
            ws.cell(row=f, column=2, value=producto)
            ws.cell(row=f, column=3, value=categoria)
            ws.cell(row=f, column=4, value=unidad)
            ws.cell(row=f, column=5, value=par)
            ws.cell(row=f, column=6, value=par_max)
            ws.cell(row=f, column=10, value=precio)
        for f in range(5, r_ult + 1):
            # RT-07 · el semáforo se REESCRIBE (la v1.1 no guardaba el par
            # level y `expandir_filas` replicaba ese defecto a las filas
            # libres).
            _f(ws, 'H{}'.format(f), F_ESTADO.format(f=f))
            _f(ws, 'I{}'.format(f), F_APEDIR.format(f=f))
            _f(ws, 'K{}'.format(f), F_VALOR.format(f=f))
        anadidas = motor.expandir_filas(ws, r_ult, r_v2, cola=(pie,),
                                        numerar=1)
        _ancho(ws, [13, 11, 22, 26], col0=10)
        cambios.append(
            '01:{}!B5:J{}: {} productos con categoría, unidad, par level/max '
            'y precio reales (DOM-01/DOM-02/DOM-14/TEC-01/TEC-25/COM-02); '
            'I y K reescritas; +{} filas libres hasta la {}'
            .format(hoja, r_ult, len(filas), anadidas, r_v2))
    _resumen_01(wb, cambios)
    _instrucciones(wb['Instrucciones'], INSTRUCCIONES_01)
    cambios.append('01:Instrucciones!A1:A24: reescritas — el «Resumen» que no '
                   'existía es «Resumen Dashboard» (§1.6) y la banda ámbar ya '
                   'propone reposición')


def _resumen_01(wb, cambios):
    """DOM-09/TEC-03/COM-14 — `Resumen Dashboard` tenía 4 cabeceras y 0
    fórmulas. Queda: 3 zonas + TOTAL arriba, y el valor del stock por cada una
    de las 10 categorías canónicas abajo."""
    ws = wb['Resumen Dashboard']
    zonas = [('Cocina', 5, 44), ('Barra', 5, 34), ('Almacén', 5, 34)]
    _nota(ws, 'A2', 'Todo sale solo de Cocina, Barra y Almacén: no se escribe '
                    'nada aquí.')
    for i, dato in enumerate(zonas):
        hoja, r0, r1 = dato
        f = 4 + i
        ws.cell(row=f, column=1, value=hoja)
        _f(ws, 'B{}'.format(f),
           '=COUNTIF(\'{h}\'!$H${a}:$H${b},"*BAJO*")'.format(h=hoja, a=r0,
                                                             b=r1))
        _f(ws, 'C{}'.format(f),
           '=COUNTIF(\'{h}\'!$H${a}:$H${b},"*PEDIR*")'.format(h=hoja, a=r0,
                                                              b=r1))
        _f(ws, 'D{}'.format(f),
           '=SUM(\'{h}\'!$K${a}:$K${b})'.format(h=hoja, a=r0, b=r1))
    ws['A7'] = 'TOTAL'
    ws['A7'].font = Font(bold=True)
    for col in ('B', 'C', 'D'):
        _f(ws, '{}7'.format(col), '=SUM(${c}$4:${c}$6)'.format(c=col))
        ws['{}7'.format(col)].font = Font(bold=True)

    ws['A8'] = 'VALOR DEL STOCK POR CATEGORÍA'
    ws['A8'].font = Font(bold=True)
    _cabecera(ws, 9, ['Categoría', 'Valor (€)', '% del Total'])
    for i, cat in enumerate(motor.CATEGORIAS):
        f = 10 + i
        ws.cell(row=f, column=1, value=cat)
        trozos = ['SUMIF(\'{h}\'!$C${a}:$C${b},$A{f},\'{h}\'!$K${a}:$K${b})'
                  .format(h=h, a=a, b=b, f=f) for h, a, b in zonas]
        _f(ws, 'B{}'.format(f), '=' + '+'.join(trozos))
        _f(ws, 'C{}'.format(f), '=IFERROR($B{f}/$B$20,"")'.format(f=f))
    ws['A20'] = 'VALOR TOTAL DEL STOCK'
    ws['A20'].font = Font(bold=True)
    _f(ws, 'B20', '=SUM($B$10:$B$19)')
    ws['B20'].font = Font(bold=True)
    _nota(ws, 'A22', 'El valor total de abajo y la suma de la columna «Valor '
                     'total stock (€)» de arriba tienen que coincidir: si no, '
                     'hay una categoría escrita a mano fuera del desplegable.')
    for f in range(4, 8):
        ws.cell(row=f, column=2).number_format = motor.FMT_ENT
        ws.cell(row=f, column=3).number_format = motor.FMT_ENT
        ws.cell(row=f, column=4).number_format = motor.FMT_EUR
    for f in range(10, 21):
        ws.cell(row=f, column=2).number_format = motor.FMT_EUR
        ws.cell(row=f, column=3).number_format = motor.FMT_PCT1
    _ancho(ws, [30, 26, 22, 22])
    cambios.append('01:Resumen Dashboard!A4:D20: 3 zonas + TOTAL (COUNTIF de '
                   'BAJO/PEDIR y SUM del valor) y las 10 categorías con '
                   'SUMIF — la hoja tenía 0 fórmulas (DOM-09/TEC-03/COM-14)')


# ==========================================================================
# 02 — fichas de proveedores
# ==========================================================================
def _pre_02(wb, cambios):
    ws = wb['Directorio Proveedores']
    if 'CIF' not in str(ws['D3'].value or ''):
        for idx in INSERTS_DIRECTORIO:
            motor.insertar_columna(ws, idx)
        cambios.append('02:Directorio Proveedores!A3:S3: +8 columnas '
                       '(CIF/NIF, Nº RGSEAA, contacto de incidencias, día de '
                       'pedido, plazo, homologación y sus dos fechas) — '
                       'DOM-17')
    ws = wb['Comparativa Precios']
    if 'Formato' not in str(ws['D3'].value or ''):
        motor.insertar_columna(ws, 4)
        motor.insertar_columna(ws, 4)
        cambios.append('02:Comparativa Precios!D3:E3: +2 columnas (formato de '
                       'venta y contenido) — sin ellas MIN compara la garrafa '
                       'de 5 L con el litro (DOM-16)')


def _post_02(wb, cambios):
    _directorio_02(wb, cambios)
    _comparativa_02(wb, cambios)
    _evaluacion_02(wb, cambios)
    _condiciones_02(wb, cambios)
    _instrucciones(wb['Instrucciones'], INSTRUCCIONES_02)
    cambios.append('02:Instrucciones!A7:A10: los nombres de las cuatro '
                   'pestañas ya casan con wb.sheetnames y los criterios de '
                   'evaluación son los cinco reales (§1.6, COM-29)')


def _directorio_02(wb, cambios):
    ws = wb['Directorio Proveedores']
    _limpiar_dv_a(ws)
    _cabecera(ws, 3, CAB_DIRECTORIO)
    _ancho(ws, ANCHOS_DIRECTORIO)
    _nota(ws, 'A2', 'Fichas de ejemplo (una por familia de compra). El Nº '
                    'RGSEAA te lo da el proveedor: pídeselo por escrito y '
                    'guárdalo con su ficha técnica.')
    for i, prov in enumerate(PROVEEDORES):
        motor.sembrar(ws, 4 + i, list(prov), col0=2, marca='(ejemplo)')
    for f in range(4, 24):
        ws.cell(row=f, column=13).number_format = motor.FMT_ENT
        ws.cell(row=f, column=14).number_format = motor.FMT_EUR
        ws.cell(row=f, column=17).number_format = motor.FMT_FECHA
        ws.cell(row=f, column=18).number_format = motor.FMT_FECHA
    _dv(ws, 'C4:C23', motor.CATEGORIAS, 'Categoría no válida',
        'La familia principal que te sirve este proveedor. Son las 10 '
        'categorías del kit.')
    _dv(ws, 'P4:P23', ['S', 'N', 'Pendiente'], 'Homologado (S/N)',
        'S sólo cuando tengas su CIF/NIF, su Nº RGSEAA y la ficha técnica '
        'de los productos que te sirve. Es el prerrequisito de homologación '
        'de proveedores de tu plan APPCC.')
    cambios.append('02:Directorio Proveedores!A4:S9: 6 fichas de ejemplo '
                   '(carnes, pescados, frutas y verduras, secos, bebidas, '
                   'limpieza) + DV de categoría y de homologación — la hoja '
                   'estaba en blanco salvo la numeración (COM-27/DOM-17)')


def _comparativa_02(wb, cambios):
    """DOM-15/DOM-16/TEC-07/COM-18.

    El bloque `P:T` (oculto) normaliza cada precio dividiéndolo por el
    contenido de la línea; MIN, INDEX/MATCH y el % corren sobre él. La guarda
    de la SPEC (`=IFERROR(F4/$E4,"")`) se ENDURECE: con `F4` vacía la división
    da 0 —no error—, así que MIN devolvería 0 y la tabla volvería a enseñar
    diez ceros, que es exactamente el defecto TEC-07.
    """
    ws = wb['Comparativa Precios']
    # Idempotencia: la DV de unidad se vuelve a añadir en cada pasada, así que
    # hay que borrar las de la pasada anterior o la 2.ª pasada las duplica.
    _limpiar_dv_a(ws)
    _cabecera(ws, 3, ['#', 'Producto', 'Unidad', 'Formato de venta',
                      'Contenido (kg/L/ud)'])
    _cabecera(ws, 3, ['Mejor precio/ud (€)', 'Mejor proveedor',
                      '% Diferencia', 'Fecha de cotización',
                      'Vigencia hasta'], col0=11)
    _cabecera(ws, 3, ['Norm. 1', 'Norm. 2', 'Norm. 3', 'Norm. 4', 'Norm. 5'],
              col0=16)
    # RD-18 · la columna U va DESPUÉS del bloque oculto P:T, así que en
    # pantalla aparece pegada a «Vigencia hasta», que es donde el usuario la
    # busca, sin tener que desplazar las cinco columnas de normalización.
    _cabecera(ws, 3, ['Estado de la cotización'], col0=21)
    _nota(ws, 'A2', 'Las cinco columnas de precio son los cinco primeros '
                    'proveedores del directorio: pídeles cotización del MISMO '
                    'formato de venta. Rellena Formato de venta y Contenido '
                    'ANTES que los precios: sin Contenido no hay precio por '
                    'unidad que comparar y la fila te lo dice. Las columnas P '
                    'a T (ocultas) son ese precio por unidad. Y mira la '
                    'columna Estado: un comparativo vencido es peor que no '
                    'tenerlo, porque negocias con un dato que el proveedor ya '
                    'no sostiene.')
    for i, prov in enumerate(PROVEEDORES[:5]):
        col = motor.get_column_letter(6 + i)
        _f(ws, '{}3'.format(col),
           '=IF(\'Directorio Proveedores\'!$B${f}="","Prov. {n}",'
           '\'Directorio Proveedores\'!$B${f})'.format(f=4 + i, n=i + 1))
        ws['{}3'.format(col)]._style = copy.copy(ws['B3']._style)
    for i, dato in enumerate(COMPARATIVA):
        f = 4 + i
        producto, unidad, formato, contenido, _ref = dato
        ws.cell(row=f, column=2, value=producto)
        ws.cell(row=f, column=3, value=unidad)
        ws.cell(row=f, column=4, value=formato)
        ws.cell(row=f, column=5, value=contenido)
        for j in range(5):
            origen = motor.get_column_letter(6 + j)
            # RT-02 · la guarda del DENOMINADOR contemplaba el 0 y la del
            # NUMERADOR no, que es justo el caso que describía el R1: un «0»
            # tecleado para decir «no me lo sirve» ganaba el MIN, «Mejor
            # precio/ud» salía 0,00 € y la comparativa señalaba como ganador
            # al proveedor que NO sirve el producto.
            _f(ws, '{}{}'.format(motor.get_column_letter(16 + j), f),
               '=IF(OR(${c}{f}="",${c}{f}<=0,$E{f}="",$E{f}<=0),"",'
               'IFERROR(${c}{f}/$E{f},""))'.format(c=origen, f=f))
        # RT-12 · si falta «Contenido» toda la comparativa se apagaba sin una
        # sola pista de por qué. El kit ya tiene doctrina para esto (§1.8, el
        # «⚠ falta coste» de 03!G, 05!G y BONUS-08!G): aquí se aplica igual.
        _f(ws, 'K{}'.format(f),
           '=IF(COUNT($F{f}:$J{f})=0,"",IF($E{f}="","⚠ falta contenido",'
           'IF(COUNT($P{f}:$T{f})=0,"",MIN($P{f}:$T{f}))))'.format(f=f))
        # La guarda `IF($K="";"")` de delante NO sobra: con la fila sin
        # cotizar, `P:T` son cinco cadenas vacías y `MATCH("";…;0)` CASA con
        # la primera —en Excel y en pycel—, así que sin ella la columna
        # señalaría al primer proveedor del directorio como el más barato de
        # una línea en la que nadie ha cotizado. Medido el 2026-08-23 en
        # `Comparativa Precios!L6`.
        _f(ws, 'L{}'.format(f),
           '=IF(OR($K{f}="",$K{f}="⚠ falta contenido"),"",'
           'IFERROR(INDEX($F$3:$J$3,MATCH($K{f},$P{f}:$T{f},0)),""))'
           .format(f=f))
        _f(ws, 'M{}'.format(f),
           '=IF(OR($K{f}="",$K{f}="⚠ falta contenido"),"",'
           'IFERROR((MAX($P{f}:$T{f})-$K{f})/$K{f},""))'.format(f=f))
        # RD-18 · la fecha de cotización y la vigencia existían y NADIE las
        # miraba: «Mejor precio», «Mejor proveedor» y «% Diferencia» daban
        # veredicto con precios caducados y sin decirlo.
        _f(ws, 'U{}'.format(f),
           '=IF($O{f}="","",IF($O{f}<TODAY(),"⛔ COTIZACIÓN VENCIDA",'
           'IF($O{f}-TODAY()<=7,"🟡 vence esta semana","🟢 vigente")))'
           .format(f=f))
        ws.cell(row=f, column=15).value = None
    for i, precios in COTIZACIONES.items():
        valores, fecha = precios
        for j, p in enumerate(valores):
            if p is not None:
                ws.cell(row=4 + i, column=6 + j, value=p)
        ws.cell(row=4 + i, column=14, value=fecha)
        ws.cell(row=4 + i, column=15,
                value=fecha + datetime.timedelta(days=30))
    for f in range(4, 14):
        for c in range(6, 12):
            ws.cell(row=f, column=c).number_format = motor.FMT_EUR
        for c in range(16, 21):
            ws.cell(row=f, column=c).number_format = motor.FMT_EUR
        ws.cell(row=f, column=13).number_format = motor.FMT_PCT1
        ws.cell(row=f, column=14).number_format = motor.FMT_FECHA
        ws.cell(row=f, column=15).number_format = motor.FMT_FECHA
    # RC-20 · la única hoja del kit sin una sola validación de datos: las
    # unidades se tecleaban libres mientras en las otras ocho salen de
    # desplegable.
    _dv(ws, 'C4:C13', motor.UNIDADES, 'Unidad no válida',
        'La misma lista de unidades que el resto del kit. El formato de venta '
        '(caja, garrafa, saco…) va en la columna de al lado.')
    _ancho(ws, [5, 28, 9, 22, 15, 13, 13, 13, 13, 13, 15, 26, 12, 14, 13,
                10, 10, 10, 10, 10, 24])
    for col in ('P', 'Q', 'R', 'S', 'T'):
        ws.column_dimensions[col].hidden = True
    cambios.append('02:Comparativa Precios!K4:M13 + U4:U13: mejor precio con '
                   'guarda de COUNT y de CERO (RT-02), aviso «⚠ falta '
                   'contenido» (RT-12), mejor PROVEEDOR con INDEX/MATCH, % de '
                   'diferencia sobre el bloque oculto P:T y ESTADO de la '
                   'cotización con su semáforo (RD-18); productos y unidades '
                   'idénticos a los del 01 y DV de unidad (RC-20); los '
                   'precios de ejemplo derivan del precio/ud de referencia '
                   'del 01 (RD-11) (DOM-15/DOM-16/TEC-07/COM-18)')


def _evaluacion_02(wb, cambios):
    """TEC-16/COM-29 — «TOTAL» que era una media, guarda de un solo criterio
    (un proveedor puntuado sólo en Calidad salía con un 5,0) y «Nota» vacía."""
    ws = wb['Evaluación Proveedores']
    ws['H3'] = 'PUNTUACIÓN MEDIA (1-5)'
    ws['I3'] = 'Nota'
    ws['J3'] = 'Coherencia con el directorio'
    ws['J3']._style = copy.copy(ws['I3']._style)
    for i, prov in enumerate(PROVEEDORES):
        f = 4 + i
        motor.sembrar(ws, f, [prov[0]] + list(EVALUACIONES[i]), col0=2,
                      marca='(ejemplo)')
    for f in range(4, 14):
        _f(ws, 'H{}'.format(f),
           '=IF(COUNT($C{f}:$G{f})<3,"",AVERAGE($C{f}:$G{f}))'.format(f=f))
        _f(ws, 'I{}'.format(f),
           '=IF($H{f}="","",IF($H{f}>=4.5,"A — Preferente",'
           'IF($H{f}>=3.5,"B — Válido",IF($H{f}>=2.5,"C — Vigilar",'
           '"D — Sustituir"))))'.format(f=f))
        # RD-19 · la fórmula que cruza NOTA con HOMOLOGACIÓN. La homologación
        # de proveedores es prerrequisito del plan APPCC: un «D — Sustituir»
        # que sigue con la S puesta en el directorio es una contradicción que
        # el propio fichero tiene que cantar.
        _f(ws, 'J{}'.format(f),
           '=IF($I{f}="","",IF(IFERROR(VLOOKUP($B{f},'
           "'Directorio Proveedores'!$B$4:$P$23,15,FALSE),\"\")<>\"S\","
           '"—",IF(LEFT($I{f},1)="D",'
           '"⚠ nota D y sigue homologado: revísalo",'
           'IF(LEFT($I{f},1)="C","🟡 nota C: vigílalo, sigue homologado",'
           '"✓ coherente"))))'.format(f=f))
        ws.cell(row=f, column=8).number_format = '0.0'
    _nota(ws, 'A2', 'La nota necesita al menos 3 criterios puntuados: con uno '
                    'solo la media no significa nada. La última columna cruza '
                    'la nota con la casilla «Homologado» del directorio: un '
                    '«D — Sustituir» que sigue homologado es una '
                    'contradicción, y la homologación de proveedores es '
                    'prerrequisito de tu plan APPCC.')
    _ancho(ws, [5, 28, 14, 14, 16, 14, 16, 20, 18, 38])
    cambios.append('02:Evaluación Proveedores!H3:J13: media con guarda de 3 '
                   'criterios, nota A/B/C/D y CRUCE nota → homologación '
                   '(RD-19); 6 evaluaciones de ejemplo que cubren las cuatro '
                   'notas y el proveedor con nota D pasa a «Pendiente» en el '
                   'directorio (TEC-16/COM-29)')


#: RD-20 · los mismos cuatro datos comerciales estaban TECLEADOS A MANO en
#: tres sitios (directorio del 02, condiciones del 02 y hoja Proveedores del
#: 03). El día que el proveedor sube el mínimo de 150 a 200 €, el usuario lo
#: cambia en uno y los otros dos siguen mintiendo. Dentro del 02 el enlace es
#: trivial y se hace; entre libros no se puede (§6 prohíbe las referencias
#: externas), así que en el 03 se dice por escrito de dónde sale la copia.
#: (columna de Condiciones, columna del Directorio)
ENLACES_CONDICIONES = [
    ('B', 'B'),      # Proveedor
    ('C', 'O'),      # Plazo de pago  ← Condiciones de pago
    ('D', 'N'),      # Pedido Mínimo (€)
    ('H', 'K'),      # Días de pedido ← Día de pedido
    ('I', 'L'),      # Días de reparto ← Día de reparto
]


def _condiciones_02(wb, cambios):
    ws = wb['Condiciones Comerciales']
    _cabecera(ws, 3, CAB_CONDICIONES)
    for i, cond in enumerate(CONDICIONES):
        # Las columnas enlazadas (B, C, D, H, I) se escriben como fórmula
        # justo debajo: aquí sólo se siembra lo que es propio de esta hoja.
        valores = list(cond)
        valores[0] = None       # C · plazo de pago  → enlace
        valores[1] = None       # D · pedido mínimo  → enlace
        valores[5] = None       # H · días de pedido → enlace
        valores[6] = None       # I · días de reparto → enlace
        motor.sembrar(ws, 4 + i, [None] + valores, col0=2)
    for f in range(4, 14):
        for destino, origen in ENLACES_CONDICIONES:
            _f(ws, '{}{}'.format(destino, f),
               "=IF('Directorio Proveedores'!${o}{f}=\"\",\"\","
               "'Directorio Proveedores'!${o}{f})".format(o=origen, f=f))
        ws.cell(row=f, column=4).number_format = motor.FMT_EUR
        ws.cell(row=f, column=5).number_format = motor.FMT_PCT1
    _nota(ws, 'A2', 'Proveedor, plazo de pago, pedido mínimo y días de pedido '
                    'y de reparto SE TRAEN SOLOS del Directorio de '
                    'Proveedores: se escriben una vez y en un solo sitio. Lo '
                    'que se rellena aquí es lo propio de la negociación: '
                    'rappel, transporte, horario de entrega y notas.')
    _ancho(ws, [5, 28, 22, 16, 12, 24, 18, 16, 16, 34])
    cambios.append('02:Condiciones Comerciales!B4:I13: proveedor, plazo de '
                   'pago, pedido mínimo y días de pedido/reparto pasan a ser '
                   'ENLACES al Directorio (RD-20) — estaban tecleados a mano '
                   'en tres sitios del kit sin una sola fórmula que los '
                   'ligara; «Rappel» pasa a «Rappel (%)» y «Pedido Mínimo» a '
                   'euros (COM-27/§1.7)')


# ==========================================================================
# BONUS-08 — inventario rápido mensual
# ==========================================================================
CAB_08 = ['Compras del mes (uds)', 'Consumo del mes (uds)',
          'Variación (uds)', '% Variación']

#: DOM-32 — el inventario mensual se hace para saber el CONSUMO:
#: stock anterior + compras − stock final. Sin la columna de compras es
#: incalculable, y era justo la que faltaba.
#: RD-15/RT-18 — la guarda vigilaba F y H y se dejaba fuera I, que es el
#: sumando que MÁS se olvida (hay que ir a buscar los albaranes del mes). Una
#: celda vacía suma 0 en Excel, así que el «consumo del mes» —que las propias
#: Instrucciones venden como «el dato con el que se calcula el food cost
#: real»— publicaba sin avisar la simple bajada de stock: un número plausible
#: y falso. Dos columnas antes, en la misma fila, el fichero ya sabía avisar
#: («⚠ falta coste»); ahora lo hace también aquí (§1.8, misma doctrina).
F_CONSUMO = ('=IF(OR($F{f}="",$H{f}=""),"",'
             'IF($I{f}="","⚠ faltan compras",$H{f}+$I{f}-$F{f}))')
#: TEC-24 — la guarda de la v1.1 miraba F (stock actual) y la fórmula dependía
#: de H (stock del mes anterior): la primera vez que se usa la plantilla toda
#: la columna presentaba el stock entero como crecimiento.
F_VARIACION = '=IF(OR($F{f}="",$H{f}=""),"",$F{f}-$H{f})'
#: RT-05 — la columna nueva guardaba el stock del mes ANTERIOR pero no el
#: ACTUAL: con el stock anterior cargado y la línea todavía sin contar
#: publicaba −100,0 % y el formato condicional la pintaba en ámbar. Mientras
#: el conteo está a medias —que es todo el rato que se usa la hoja— las
#: líneas sin contar aparecían como caídas del 100 %. Sus dos hermanas de la
#: misma fila (J y K) sí guardaban $F: era una inconsistencia, no un criterio.
F_PCT = '=IF(OR($F{f}="",$H{f}="",$H{f}=0),"",$F{f}/$H{f}-1)'

R_08_ULT, R_08_V2, R_08_PIE = 54, 84, 56


def _post_08(wb, cambios):
    ws = wb['Conteo Rápido']
    _cabecera(ws, 4, CAB_08, col0=9)
    _nota(ws, 'A3', NOTA_EJEMPLO + ' Son los mismos 50 productos de la '
                    'plantilla 01.')
    todos = []
    for hoja, filas in PRODUCTOS:
        todos += filas
    for i, dato in enumerate(todos):
        f = 5 + i
        producto, categoria, unidad, par, par_max, precio = dato
        ws.cell(row=f, column=2, value=producto)
        ws.cell(row=f, column=3, value=categoria)
        ws.cell(row=f, column=4, value=unidad)
        ws.cell(row=f, column=5, value=precio)
    for f in range(5, R_08_ULT + 1):
        _f(ws, 'G{}'.format(f), motor.guarda_doble('F', 'E', f))
        _f(ws, 'J{}'.format(f), F_CONSUMO.format(f=f))
        _f(ws, 'K{}'.format(f), F_VARIACION.format(f=f))
        _f(ws, 'L{}'.format(f), F_PCT.format(f=f))
        ws.cell(row=f, column=9).value = None
    anadidas = motor.expandir_filas(ws, R_08_ULT, R_08_V2, cola=(R_08_PIE,),
                                    numerar=1)
    # La fila del TOTAL queda por DEBAJO del bloque de datos, así que
    # `motor.aplicar_formatos` —que sólo recorre el bloque— no la alcanza y
    # el total del stock se imprimiría sin el símbolo del euro (§1.7).
    fila_total = R_08_PIE + (R_08_V2 - R_08_ULT)
    ws.cell(row=fila_total, column=6).value = 'TOTAL VALORADO:'
    ws.cell(row=fila_total, column=7).number_format = motor.FMT_EUR
    _ancho(ws, [5, 30, 22, 10, 14, 13, 14, 17, 17, 18, 14, 12])
    ref = 'L5:L{}'.format(R_08_V2)
    motor._limpiar_cf(ws, set([ref]))
    ws.conditional_formatting.add(ref, Rule(
        type='expression', dxf=motor._dxf('ambar'), stopIfTrue=True,
        formula=['AND($L5<>"",ABS($L5)>0.2)']))
    cambios.append(
        'BONUS-08:Conteo Rápido!I4:L{}: «Compras del mes», «Consumo del mes» '
        '(stock anterior + compras − stock final), variación con guarda de H '
        'y % de variación con ámbar por encima del ±20 %; G con doble guarda '
        'de cantidad y precio; 50 productos coherentes con el 01; +{} filas '
        'hasta la {} (DOM-32/TEC-24/COM-28, §1.8)'
        .format(R_08_V2, anadidas, R_08_V2))
    _instrucciones(wb['Instrucciones'], INSTRUCCIONES_08)


# ==========================================================================
# Instrucciones (§1.10) — describen lo que HAY
# ==========================================================================
#: Sin comillas simples alrededor de nombres: el gate §1.6
#: (`motor.pestanas_citadas`) sólo tolera lo entrecomillado si es una pestaña
#: real o una cabecera de columna, y una comilla de más es un falso rojo.
INSTRUCCIONES_01 = [
    'Inventario de Stock Diario',
    'AI Chef Pro — aichef.pro',
    None,
    'Controla el stock diario de tu restaurante: qué tienes, qué te falta y '
    'cuánto vale.',
    'Cada pestaña es una zona de conteo (Cocina, Barra y Almacén) y la '
    'pestaña Resumen Dashboard las totaliza.',
    None,
    'CÓMO USAR:',
    '1. Revisa los productos de ejemplo y sustitúyelos por los tuyos. Las '
    'columnas Categoría y Unidad tienen desplegable: úsalo, porque los '
    'análisis del kit agregan por ese texto.',
    '2. Ajusta Par Level (stock mínimo de trabajo) y Par Max (hasta dónde '
    'repones) producto a producto.',
    '3. Escribe el precio de compra SIN IVA en Precio/ud (€): es lo que hace '
    'que la columna Valor (€) y el resumen funcionen.',
    '4. Cada día anota lo que cuentas en Stock Real.',
    '5. La columna Estado se colorea sola:',
    '   🔴 PEDIR = por debajo del par level',
    '   🟡 BAJO = entre el par level y una vez y media el par level',
    '   🟢 OK = por encima',
    '6. A Pedir calcula cuánto reponer hasta el Par Max ya en ámbar, no sólo '
    'en rojo: si esperas al rojo, con dos días de plazo de entrega llegas '
    'tarde.',
    '7. Estado y A Pedir se quedan EN BLANCO mientras la fila no tenga par '
    'level: en las filas libres del final no hay ninguno, y decirte «OK» '
    'comparando contra una casilla vacía sería mentirte justo en el producto '
    'que acabas de dar de alta.',
    None,
    'PERSONALIZACIÓN:',
    '- Hay filas libres al final de cada zona: Cocina hasta la 44, Barra y '
    'Almacén hasta la 34.',
    '- EL PAR LEVEL ES TU PUNTO DE PEDIDO. Sale de: consumo diario × '
    '(plazo de entrega + días de cobertura de seguridad). Es exactamente la '
    'misma cuenta que hace el BONUS-09, así que los dos ficheros dan el '
    'mismo número para el mismo producto: no hay dos reglas, hay una.',
    '- Y el Par Max es el techo: la cantidad sugerida del BONUS-09 nunca lo '
    'supera, porque de nada sirve que la fórmula mande pedir más de lo que '
    'te cabe en cámara.',
    '- Una misma referencia puede aparecer en DOS zonas (las servilletas '
    'están en Barra y en Almacén): es stock por ubicación y cada zona lleva '
    'su propio par level. El resumen las cuenta por separado a propósito; el '
    'pedido se consolida en la plantilla 03, que es donde cuentan el pedido '
    'mínimo, los portes y el rappel.',
    '- La pestaña Resumen Dashboard totaliza el stock bajo, los productos a '
    'pedir y el valor del stock por zona y por categoría.',
    '- Los 50 productos, con sus categorías, unidades, par levels y precios, '
    'son un ejemplo real de restaurante: revísalos con tus datos antes de '
    'usarlos.',
    None,
    motor.VERSION_LINE,
]

INSTRUCCIONES_02 = [
    'Fichas de Proveedores',
    'AI Chef Pro — aichef.pro',
    None,
    'Da de alta a tus proveedores, homológalos y compara sus precios en un '
    'solo archivo.',
    None,
    'PESTAÑAS:',
    '1. Directorio Proveedores — datos fiscales (CIF/NIF), sanitarios '
    '(Nº RGSEAA), contacto, días de pedido y de reparto, y homologación',
    '2. Comparativa Precios — hasta 5 proveedores por producto, comparados '
    'por precio normalizado por unidad',
    '3. Evaluación Proveedores — calidad, precio, puntualidad, servicio y '
    'flexibilidad (1-5), con nota A/B/C/D',
    '4. Condiciones Comerciales — plazo de pago, pedido mínimo, rappel, '
    'transporte y horarios',
    None,
    'CÓMO USAR:',
    '1. Da de alta cada proveedor en el directorio. Pídele por escrito su '
    'CIF/NIF y su Nº RGSEAA: sin eso no puedes homologarlo ni darlo de alta '
    'en contabilidad.',
    '2. Marca Homologado y anota la fecha. Revisa la documentación al menos '
    'una vez al año: es el prerrequisito de homologación de proveedores que '
    'te pide tu plan APPCC.',
    '3. En la comparativa rellena el formato de venta y el contenido de cada '
    'línea ANTES de meter precios. Es lo que evita comparar la garrafa de 5 L '
    'con el litro y elegir al proveedor más caro creyendo que es el barato.',
    '4. Escribe el precio de cada proveedor en su columna. El mejor precio, '
    'el mejor proveedor y el % de diferencia salen solos.',
    '5. Puntúa a tus proveedores cada trimestre. La nota necesita al menos 3 '
    'criterios puntuados.',
    None,
    'Las 6 fichas de ejemplo están marcadas (ejemplo): bórralas cuando metas '
    'las tuyas.',
    None,
    motor.VERSION_LINE,
]

INSTRUCCIONES_08 = [
    'BONUS: Inventario Rápido Mensual',
    'AI Chef Pro — aichef.pro',
    None,
    'Hoja simplificada para el inventario mensual completo. Trae los mismos '
    '50 productos, categorías, unidades y precios de referencia que la '
    'plantilla 01, para que los dos inventarios hablen el mismo idioma.',
    None,
    'CÓMO USAR:',
    '1. Imprímela y recorre el almacén contando. Anota lo que cuentas en '
    'Stock Actual.',
    '2. Revisa el Precio/ud (€) de cada línea: es lo que valora tu stock. Si '
    'falta, la fila avisa en vez de valorarla a cero.',
    '3. Copia el Stock Mes Anterior del inventario del mes pasado y anota las '
    'compras del mes en unidades.',
    '4. El consumo del mes sale solo: stock anterior + compras − stock '
    'actual. Es el dato con el que se calcula el food cost real, y el que no '
    'te da una simple foto del stock. Si te dejas las compras, la casilla '
    'dice «faltan compras» en vez de darte la simple bajada de stock: sería '
    'un número plausible y falso, y el que peor se detecta.',
    '5. La variación y el % de variación comparan con el mes anterior; el '
    'ámbar marca las desviaciones de más del 20 %, que son las que hay que '
    'explicar. Mientras no hayas contado la línea, el % se queda EN BLANCO: '
    'una línea sin contar no es una caída del 100 %.',
    None,
    '- El consumo y la variación van en la unidad de cada producto: no los '
    'sumes entre sí, mezclarían kilos con litros. El total en euros es el de '
    'la columna Valor (€).',
    '- Hay filas libres hasta la 84.',
    None,
    motor.VERSION_LINE,
]


# ==========================================================================
# Entradas del pipeline
# ==========================================================================
def pre(wb, fname, cambios):
    """Sólo inserción de COLUMNAS: corre antes de que el motor fije rangos."""
    if fname.startswith('01-'):
        _pre_01(wb, cambios)
    elif fname.startswith('02-'):
        _pre_02(wb, cambios)


def post(wb, fname, cambios, registro=None):
    if fname.startswith('01-'):
        _post_01(wb, cambios)
    elif fname.startswith('02-'):
        _post_02(wb, cambios)
    elif fname.startswith('BONUS-08'):
        _post_08(wb, cambios)


# ==========================================================================
# Demostraciones con pycel (SPEC §5): se cambian INPUTS y se comprueba que el
# resultado se mueve en la dirección correcta. Cada bloque cita fichero:hoja:
# celda.
# ==========================================================================
def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            v = xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)
    return round(v, 4) if isinstance(v, float) else v


def _set(xl, ref, valor):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            xl.evaluate(ref)
            xl.set_value(ref, valor)
            return True
        except Exception:                                        # noqa: BLE001
            return False


def _compilar(path):
    from pycel import ExcelCompiler
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        return ExcelCompiler(filename=path)


def _demo_valoracion(path):
    """01 — DOM-09/TEC-03/COM-14: el valor del stock y el dashboard se mueven
    con la cantidad contada. En la v1.1 la columna no tenía fórmula."""
    xl = _compilar(path)
    pruebas = []
    for etiqueta, stock in (('sin contar', ''), ('10 kg contados', 10),
                            ('4 kg contados', 4)):
        _set(xl, "'Cocina'!G5", stock)
        pruebas.append({
            'caso': etiqueta, 'Cocina!G5': stock,
            'Cocina!J5_precio': _ev(xl, "'Cocina'!J5"),
            'Cocina!K5_valor': _ev(xl, "'Cocina'!K5"),
            'Resumen!D4_valor_cocina': _ev(xl, "'Resumen Dashboard'!D4"),
        })
    return {
        'ref': '01-inventario-stock-diario.xlsx:Cocina:K5 + '
               'Resumen Dashboard:D4',
        'formula': 'K5 ' + F_VALOR.format(f=5),
        'pruebas': pruebas,
        'ok': (pruebas[1]['Cocina!K5_valor'] == 65.0
               and pruebas[2]['Cocina!K5_valor'] == 26.0),
        'nota': 'en la v1.1 la columna Valor (€) no tenía fórmula en ninguna '
                'de las 50 filas y el Resumen Dashboard tenía 0 fórmulas',
    }


def _demo_a_pedir(path):
    """01 — DOM-33/TEC-21: la banda ÁMBAR ya propone reposición. Con
    E5=8 / F5=20, un stock de 10 daba «BAJO» y «A Pedir 0»."""
    xl = _compilar(path)
    pruebas = []
    for etiqueta, stock in (('por debajo del par (6)', 6),
                            ('en ámbar, 8 <= 10 < 12', 10),
                            ('holgado (15)', 15)):
        _set(xl, "'Cocina'!G5", stock)
        pruebas.append({'caso': etiqueta, 'stock': stock,
                        'H5_estado': _ev(xl, "'Cocina'!H5"),
                        'I5_a_pedir': _ev(xl, "'Cocina'!I5")})
    return {
        'ref': '01-inventario-stock-diario.xlsx:Cocina:I5',
        'formula': F_APEDIR.format(f=5),
        'par_level_E5': _ev(xl, "'Cocina'!E5"),
        'par_max_F5': _ev(xl, "'Cocina'!F5"),
        'pruebas': pruebas,
        'ok': pruebas[1]['I5_a_pedir'] == 10 and pruebas[2]['I5_a_pedir'] == 0,
        'nota': 'la v1.1 devolvía 0 en todo el tramo ámbar: el semáforo '
                'avisaba y la columna de al lado decía que no pidieras nada',
    }


def _demo_resumen_categoria(path):
    """01 — el SUMIF por categoría cruza las tres zonas. Cárnicos = pollo
    (6,50 €/kg) + solomillo (32,00 €/kg)."""
    xl = _compilar(path)

    def _leer():
        return {'B10_carnicos': _ev(xl, "'Resumen Dashboard'!B10"),
                'B20_total': _ev(xl, "'Resumen Dashboard'!B20"),
                'C10_porcentaje': _ev(xl, "'Resumen Dashboard'!C10")}

    # Lectura EN FRÍO obligatoria antes del primer `set_value`: `inject_cache`
    # deja el valor cacheado dentro del xlsx y pycel lo devuelve tal cual
    # hasta que una dependencia se marca sucia. Sin esta llamada la primera
    # medición sale 0,00 € y parece que el SUMIF no suma.
    partida = _leer()
    _set(xl, "'Cocina'!G5", 10)
    solo_pollo = _leer()
    _set(xl, "'Cocina'!G6", 2)
    antes = _leer()
    _set(xl, "'Cocina'!G6", 4)
    despues = _leer()
    return {
        'ref': '01-inventario-stock-diario.xlsx:Resumen Dashboard:B10/B20',
        'formula': 'B10 = SUMIF por categoría sobre Cocina+Barra+Almacén',
        'fichero_recien_descargado': partida,
        'con_10kg_de_pollo': solo_pollo,
        'con_10kg_pollo_y_2kg_solomillo': antes,
        'subiendo_el_solomillo_a_4kg': despues,
        'ok': (partida['B10_carnicos'] == 0
               and solo_pollo['B10_carnicos'] == 65.0
               and antes['B10_carnicos'] == 129.0
               and despues['B10_carnicos'] == 193.0),
        'nota': '10 × 6,50 + 2 × 32,00 = 129,00 € → 10 × 6,50 + 4 × 32,00 = '
                '193,00 €',
    }


def _demo_comparativa(path):
    """02 — DOM-16/TEC-07: el mejor precio se decide sobre el precio
    NORMALIZADO por unidad, no sobre el importe del formato.

    Ronda 2 · se añaden el CERO (RT-02: un «0» tecleado para decir «no me lo
    sirve» ganaba el MIN y la comparativa señalaba como ganador al proveedor
    que no sirve el producto) y la fila SIN contenido (RT-12: toda la
    comparativa se apagaba sin una sola pista de por qué). Y se comprueba que
    el mejor precio/ud coincide con el precio de referencia del 01 (RD-11: el
    AOVE costaba 5,80 €/L en el 01 y 8,40 €/L aquí).
    """
    xl = _compilar(path)
    fila = 7                      # Aceite de oliva virgen extra, garrafa 5 L
    base = {
        'B7_producto': _ev(xl, "'Comparativa Precios'!B7"),
        'E7_contenido': _ev(xl, "'Comparativa Precios'!E7"),
        'precios_F7_J7': [_ev(xl, "'Comparativa Precios'!{}7".format(c))
                          for c in 'FGHIJ'],
        'normalizados_P7_T7': [_ev(xl, "'Comparativa Precios'!{}7".format(c))
                               for c in 'PQRST'],
        'K7_mejor_precio_ud': _ev(xl, "'Comparativa Precios'!K7"),
        'L7_mejor_proveedor': _ev(xl, "'Comparativa Precios'!L7"),
        'M7_diferencia': _ev(xl, "'Comparativa Precios'!M7"),
        'U7_estado_cotizacion': _ev(xl, "'Comparativa Precios'!U7"),
        'precio_ud_del_01': 5.80,
    }
    # RT-02 · el CERO. La columna I es la que trae el mejor precio (29,00 €
    # de garrafa = 5,80 €/L); ponerle un 0 a OTRO proveedor no puede
    # convertirlo en el ganador.
    _set(xl, "'Comparativa Precios'!H7", 0)
    cero = {'H7': 0,
            'K7_mejor_precio_ud': _ev(xl, "'Comparativa Precios'!K7"),
            'L7_mejor_proveedor': _ev(xl, "'Comparativa Precios'!L7")}
    _set(xl, "'Comparativa Precios'!H7", 30.20)
    # el ganador de verdad sube: la comparativa tiene que cambiar de nombre
    _set(xl, "'Comparativa Precios'!I7", 50.0)
    tras = {'I7': 50.0,
            'K7_mejor_precio_ud': _ev(xl, "'Comparativa Precios'!K7"),
            'L7_mejor_proveedor': _ev(xl, "'Comparativa Precios'!L7")}
    _set(xl, "'Comparativa Precios'!I7", 29.00)
    vacia = {'fila': 6,
             'K6_mejor_precio_ud': _ev(xl, "'Comparativa Precios'!K6"),
             'L6_mejor_proveedor': _ev(xl, "'Comparativa Precios'!L6")}
    # RT-12 · una cotización SIN contenido: en vez de apagarse en silencio,
    # la casilla dice qué falta (la doctrina «⚠ falta coste» de §1.8).
    _set(xl, "'Comparativa Precios'!F6", 40.0)
    _set(xl, "'Comparativa Precios'!E6", '')
    sin_contenido = {'fila': 6, 'F6': 40.0, 'E6': '(vacío)',
                     'K6_mejor_precio_ud': _ev(xl, "'Comparativa Precios'!K6"),
                     'L6_mejor_proveedor': _ev(xl, "'Comparativa Precios'!L6")}
    return {
        'ref': '02-fichas-proveedores.xlsx:Comparativa Precios:K7/L7/M7/U7',
        'formula': 'K = IF(COUNT(F:J)=0;"";IF(E="";"⚠ falta contenido";'
                   'MIN(P:T))) sobre P = precio / contenido, con P vacío si '
                   'el precio es 0 o negativo',
        'con_las_cotizaciones_de_ejemplo': base,
        'con_un_cero_tecleado_en_otro_proveedor': cero,
        'subiendo_al_mas_barato_a_50': tras,
        'fila_sin_cotizaciones': vacia,
        'fila_con_precio_y_sin_contenido': sin_contenido,
        'ok': (base['K7_mejor_precio_ud'] == 5.8
               and base['L7_mejor_proveedor']
               == motor.PROVEEDORES_MARCADOS[3]
               and cero['K7_mejor_precio_ud'] == 5.8
               and cero['L7_mejor_proveedor']
               == motor.PROVEEDORES_MARCADOS[3]
               and tras['K7_mejor_precio_ud'] != 5.8
               and tras['L7_mejor_proveedor']
               != motor.PROVEEDORES_MARCADOS[3]
               and vacia['K6_mejor_precio_ud'] in ('', None)
               and sin_contenido['K6_mejor_precio_ud']
               == '⚠ falta contenido'
               and base['U7_estado_cotizacion'] not in ('', None)),
        'nota': 'la v1.1 hacía MIN sobre el importe y sin guarda de vacío: '
                'las 10 filas enseñaban un 0 nada más abrir el fichero '
                '(TEC-07) y la columna Mejor Proveedor estaba vacía (DOM-15). '
                'La v2.0 lo arregló a medias: la guarda excluía la celda '
                'VACÍA pero no el CERO (RT-02) y una fila sin contenido '
                'apagaba la comparativa sin decir por qué (RT-12). Y el '
                'mejor precio/ud tiene que ser EXACTAMENTE el precio de '
                'referencia del 01 (RD-11)',
    }


def _demo_evaluacion(path):
    """02 — TEC-16: la media exige 3 criterios y la nota A/B/C/D existe."""
    xl = _compilar(path)
    pruebas = [{'caso': 'ejemplo con los 5 criterios (fila 4)',
                'H4': _ev(xl, "'Evaluación Proveedores'!H4"),
                'I4': _ev(xl, "'Evaluación Proveedores'!I4")}]
    for etiqueta, valores in (
            ('sólo Calidad puntuada', (5, '', '', '', '')),
            ('proveedor flojo', (2, 3, 2, 2, 2)),
            ('proveedor preferente', (5, 5, 5, 4, 5))):
        for col, v in zip('CDEFG', valores):
            _set(xl, "'Evaluación Proveedores'!{}4".format(col), v)
        pruebas.append({'caso': etiqueta, 'valores': list(valores),
                        'H4': _ev(xl, "'Evaluación Proveedores'!H4"),
                        'I4': _ev(xl, "'Evaluación Proveedores'!I4")})
    return {
        'ref': '02-fichas-proveedores.xlsx:Evaluación Proveedores:H4/I4',
        'formula': 'H4 = IF(COUNT(C4:G4)<3;"";AVERAGE(C4:G4))',
        'pruebas': pruebas,
        'ok': (pruebas[1]['H4'] in ('', None)
               and str(pruebas[2]['I4']).startswith('D')
               and str(pruebas[3]['I4']).startswith('A')),
        'nota': 'la v1.1 daba 5,0 —nota perfecta— a un proveedor puntuado '
                'sólo en Calidad, y la columna Nota estaba vacía en las 10 '
                'filas',
    }


def _demo_consumo(path):
    """BONUS-08 — DOM-32/TEC-24: consumo del mes, variación con guarda y
    doble guarda del valor."""
    xl = _compilar(path)
    pruebas = []
    for etiqueta, actual, anterior, compras in (
            ('primer inventario: sin mes anterior', 15, '', ''),
            ('anterior 20, compras 50, actual 15', 15, 20, 50),
            ('anterior 20, compras 50, actual 25', 25, 20, 50)):
        _set(xl, "'Conteo Rápido'!F5", actual)
        _set(xl, "'Conteo Rápido'!H5", anterior)
        _set(xl, "'Conteo Rápido'!I5", compras)
        pruebas.append({'caso': etiqueta,
                        'F5_actual': actual, 'H5_anterior': anterior,
                        'I5_compras': compras,
                        'G5_valor': _ev(xl, "'Conteo Rápido'!G5"),
                        'J5_consumo': _ev(xl, "'Conteo Rápido'!J5"),
                        'K5_variacion': _ev(xl, "'Conteo Rápido'!K5"),
                        'L5_pct': _ev(xl, "'Conteo Rápido'!L5")})
    _set(xl, "'Conteo Rápido'!E5", '')
    sin_precio = {'caso': 'cantidad contada y precio en blanco',
                  'G5_valor': _ev(xl, "'Conteo Rápido'!G5")}
    return {
        'ref': 'BONUS-08-inventario-rapido-mensual.xlsx:Conteo Rápido:'
               'J5/K5/L5/G5',
        'formula': 'J5 ' + F_CONSUMO.format(f=5),
        'pruebas': pruebas,
        'doble_guarda_G5': sin_precio,
        'ok': (pruebas[0]['K5_variacion'] in ('', None)
               and pruebas[1]['J5_consumo'] == 55
               and pruebas[2]['J5_consumo'] == 45
               and pruebas[2]['L5_pct'] == 0.25
               and 'falta coste' in str(sin_precio['G5_valor'])),
        'nota': 'la v1.1 no tenía columna de compras (el consumo era '
                'incalculable), la variación miraba la guarda equivocada y el '
                'valor de una línea sin precio se sumaba como 0,00 €',
    }


def demos(carpeta, origen=None):
    """Bloques de demostración de este grupo. `main.py` los mezcla con los
    suyos en el informe."""
    p01 = os.path.join(carpeta, FICHEROS[0])
    p02 = os.path.join(carpeta, FICHEROS[1])
    p08 = os.path.join(carpeta, FICHEROS[2])
    fuera = {}
    for clave, path, fn in (
            ('a_valoracion_stock_01', p01, _demo_valoracion),
            ('a_a_pedir_ambar_01', p01, _demo_a_pedir),
            ('a_resumen_por_categoria_01', p01, _demo_resumen_categoria),
            ('a_comparativa_normalizada_02', p02, _demo_comparativa),
            ('a_evaluacion_nota_02', p02, _demo_evaluacion),
            ('a_consumo_mensual_08', p08, _demo_consumo)):
        if not os.path.isfile(path):
            continue
        try:
            fuera[clave] = fn(path)
        except Exception as e:                                   # noqa: BLE001
            fuera[clave] = {'error': '{}: {}'.format(type(e).__name__, e)}
    return fuera
