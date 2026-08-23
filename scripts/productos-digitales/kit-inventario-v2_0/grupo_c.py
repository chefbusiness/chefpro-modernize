#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
grupo_c.py — §4 de `kit-inventario-v2-SPEC.md`: caducidades, costes y punto de
pedido.

Ficheros: `06-fifo-caducidades.xlsx`, `07-analisis-costes-compras.xlsx` y
`BONUS-09-calculadora-punto-pedido.xlsx`.

Qué arregla (ids del R1, `auditorias/kit-inventario-R1.json`):

  06  DOM-13/TEC-11/COM-09  el semáforo metía en el mismo cubo «🔴 URGENTE» lo
      que caduca mañana y lo que caducó hace tres días, y el protocolo de la
      hoja de al lado mandaba «usar HOY» todo lo rojo: el fichero autorizaba
      por escrito a servir producto caducado. Cuarto estado ⛔ CADUCADO —
      RETIRAR, quinto para el consumo preferente vencido, y protocolo nuevo.
      DOM-26  no distinguía «fecha de caducidad» de «consumo preferente»
      (Reg. (UE) 1169/2011, art. 24): columna «Tipo de fecha» con desplegable.
      DOM-24/COM-26  la hoja se vende como FIFO y lo que calcula es FEFO; la
      columna «Fecha Entrada» no aparecía en ninguna de las 100 fórmulas del
      libro. Ahora da «Días en almacén» y las Instrucciones explican los dos
      criterios.
      DOM-25/TEC-12  no había cantidad, ni unidad, ni valor: 200 g de perejil
      y 8 kg de solomillo generaban la misma alerta.
      DOM-27  «Alertas Caducidad» eran seis líneas de texto pidiendo filtrar a
      mano la otra hoja, sin autofiltro en ninguna hoja del kit.
      DOM-28/COM-15  el mapa de almacén estaba VACÍO y su cabecera metía carne
      cruda y lácteos en la misma cámara.
  07  DOM-10/TEC-08/COM-11/COM-23  «Top 20 Productos» y «Dashboard KPIs» no
      tenían ni una fórmula; el dashboard DESCRIBÍA EN TEXTO cómo calcular
      cada KPI y no existía celda donde meter ventas ni cubiertos. Food cost
      sobre CONSUMO (existencias + compras − existencias), coste por cubierto
      sin Limpieza ni Otros y objetivo relativo al ticket medio.
      DOM-29  «Variación vs Anterior» era incalculable: no había «Precio
      anterior».
      DOM-30  en ninguna parte se advertía de que los importes van SIN IVA.
      DOM-31  «% del Total» vacío con el denominador ya calculado al lado.
      §1.6  «Evolución Mensual» se citaba en Instrucciones y no existía.
  B09 DOM-06/TEC-02/COM-01  la EOQ llevaba el coste de pedido (2) y el de
      almacenamiento (0,5) HARDCODEADOS en las 30 fórmulas, sin columna de
      precio y sin capar por vida útil: mandaba comprar 121 kg de pescado y
      volver a pedir en 24 días mientras la hoja «Parámetros» del mismo
      fichero decía «entrega diaria obligatoria».
      DOM-12/TEC-14  el stock de seguridad se pedía en UNIDADES en la
      Calculadora y en DÍAS en «Parámetros»: quien seguía la guía del propio
      fichero se quedaba a la mitad del punto de pedido.
      TEC-15  «Frecuencia Pedido» devolvía #DIV/0! con consumo 0.
      §7-bis.4  se retira la promesa del «simulador».

Contrato con `main.py`: `FICHEROS`, `pre(wb, fname, cambios)`,
`post(wb, fname, cambios, registro)` y `demos(carpeta, origen)`. Las columnas
se insertan en `pre()` (antes de que el motor fije rangos); las filas se
AÑADEN en `post()`. Toda fórmula pasa por `motor._reg` para que `main.py`
verifique su valor cacheado.

DECISIONES DOCUMENTADAS (desviaciones mínimas del literal de la SPEC, todas
por seguridad de cálculo; el motivo va en el informe JSON):

  * `06!J` lleva una guarda más que la de la SPEC —`IF($G5="","",…)`—. Sin
    ella, una fila vacía devuelve 0 (que Excel lee como 00/01/1900), `K` da
    −46.000 y las 50 filas en blanco se pintan de «⛔ CADUCADO» en rojo.
  * `BONUS-09!H` exige también `$F` y `BONUS-09!L` también `$J`: sumar o
    multiplicar contra una celda que devuelve "" da #¡VALOR! en Excel, no 0.
  * `07!'Dashboard KPIs'!B3` pasa de «Valor Actual» a «Resultado» y
    `BONUS-09!'Parámetros'!B3` de «Valor Típico» a «Rango habitual»: la
    palabra «valor» es clave de MONEDA en `motor.FORMATO_POR_CABECERA` y
    formateaba como euros una columna de porcentajes y otra de texto.
  * `07!'Dashboard KPIs'!A8` deja de ser «Nº Proveedores Activos» —el libro
    no tiene un solo dato de proveedores y los enlaces entre libros están
    descartados (§6)— y pasa a «Productos con subida de precio > 5 %», que sí
    es calculable y es justo lo que promete la landing.
  * `06!'Control FIFO'` NO se siembra con lotes de ejemplo: sus fechas se
    quedarían congeladas en la fecha del build y en tres meses el fichero se
    entregaría con todas las líneas en «⛔ CADUCADO». Los ejemplos vivos son
    las demostraciones con pycel de este módulo.

Python 3.7 / openpyxl 3.1.3: sin walrus ni f-strings de depuración.
"""
import contextlib
import copy
import datetime
import os

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import motor

FICHEROS = [
    '06-fifo-caducidades.xlsx',
    '07-analisis-costes-compras.xlsx',
    'BONUS-09-calculadora-punto-pedido.xlsx',
]

#: Marca de las validaciones de ESTE grupo. NO puede empezar por
#: `motor.MARCA_DV` ('kitinv-v2'): `motor._limpiar_dv()` corre dentro de
#: `cerrar()` —después de `post()`— y borraría todo lo que lleve esa marca.
MARCA_C = 'kitinv-c'

NOTA_EJEMPLO = ('Datos de ejemplo; precios orientativos SIN IVA, edítalos con '
                'los tuyos.')


# ==========================================================================
# Ajustes que este grupo necesita del motor (§1.7 y verde de edición)
# ==========================================================================
def _extender_motor():
    """Dos retoques al motor, los dos con una razón medida detrás.

    1. «Tipo de fecha» es una columna de TEXTO (Caducidad / Consumo
       preferente) y contiene la palabra «fecha», que en
       `motor.FORMATO_POR_CABECERA` es clave de FECHA: sin esta regla las 50
       celdas del desplegable saldrían con formato `dd/mm/yyyy`.
    2. `07!'Dashboard KPIs'` lleva fórmulas en `B4:B9` y entradas verdes en
       `B12:B17`, en la MISMA columna. `motor.aplicar_verde()` decide por
       columna entera: al ver una fórmula en B marca la columna como
       calculada y DESPINTA el verde de las seis casillas de entrada. Se
       añade a `SIN_VERDE_AUTO` (que es exactamente para esto: «los grupos
       marcan ahí las celdas concretas con `marcar_verde()`»).
    """
    regla = ('@', ('tipo de fecha',))
    if regla not in motor.FORMATO_POR_CABECERA:
        motor.FORMATO_POR_CABECERA.insert(0, regla)
    if 'Dashboard KPIs' not in motor.SIN_VERDE_AUTO:
        motor.SIN_VERDE_AUTO = frozenset(motor.SIN_VERDE_AUTO) | {
            'Dashboard KPIs'}


_extender_motor()


# ==========================================================================
# Utilidades del grupo (mismas convenciones que grupo_a/grupo_b)
# ==========================================================================
def _f(ws, coord, formula):
    """Escribe una fórmula y la registra para que `main.py` compruebe que
    quedó con valor cacheado."""
    ws[coord] = motor._reg(ws, coord, formula)
    return formula


def _limpiar_dv_c(ws):
    """Quita las DV de ESTE grupo antes de reescribirlas. Sin esto la 2.ª
    pasada acumularía una copia y la idempotencia daría «cambia dv»."""
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if not (getattr(dv, 'promptTitle', None) or '').startswith(MARCA_C)]


def _dv(ws, ref, valores, titulo, prompt):
    """DV de lista INLINE. Aborta si no cupiera en los 255 caracteres de Excel
    en vez de escribir una lista truncada."""
    formula = '"{}"'.format(','.join(valores))
    if len(formula) > 255:
        raise ValueError('DV inline de {} caracteres (>255): {}'
                         .format(len(formula), titulo))
    dv = DataValidation(
        type='list', formula1=formula, allow_blank=True,
        showErrorMessage=True, errorStyle='stop', errorTitle=titulo,
        error='Elige un valor de la lista.',
        showInputMessage=True,
        promptTitle='{} · {}'.format(MARCA_C, titulo), prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(ref)
    return dv


def _cabecera(ws, fila, textos, col0=1):
    """Reescribe una fila de cabecera. Si la celda de referencia ya tiene la
    cabecera oscura del kit se clona su estilo; si no (hoja nueva), se pinta."""
    base = ws.cell(row=fila, column=col0)
    tiene = (motor._relleno(base) or '').endswith(motor.CAB)
    modelo = base._style if tiene else None
    for i, txt in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i)
        cel.value = txt
        if modelo is not None:
            cel._style = copy.copy(modelo)
        else:
            cel.fill = PatternFill('solid', fgColor=motor.CAB)
            cel.font = Font(bold=True, color='FFFFFF', size=10)
            cel.alignment = Alignment(horizontal='center', vertical='center',
                                      wrap_text=True)


def _nota(ws, coord, texto):
    cel = ws[coord]
    cel.value = texto
    cel.font = Font(italic=True, size=9, color='666666')
    cel.alignment = Alignment(vertical='center')


def _titulo(ws, coord, texto):
    cel = ws[coord]
    cel.value = texto
    cel.font = Font(bold=True, size=11, color='2D2D2D')


def _instrucciones(ws, lineas):
    """Reescribe `Instrucciones` de arriba abajo.

    IDEMPOTENCIA: la última línea es `motor.VERSION_LINE`, que es donde
    `motor.bio_y_version()` —que corre después, en `cerrar()`— ancla su bloque
    de tres líneas. Lo que NO se puede hacer es APPEND al final:
    `bio_y_version` escribiría encima.
    """
    modelo = ws.cell(row=4, column=1)._style
    maximo = ws.max_row
    for i, txt in enumerate(lineas):
        cel = ws.cell(row=i + 1, column=1)
        if cel.value is None and txt is not None:
            cel._style = copy.copy(modelo)
        cel.value = txt
    for r in range(len(lineas) + 1, maximo + 1):
        ws.cell(row=r, column=1).value = None


def _ancho(ws, anchos, col0=1):
    for i, a in enumerate(anchos):
        ws.column_dimensions[motor.get_column_letter(col0 + i)].width = a


def _clonar_estilo(ws, columnas, r0, r1, modelo=2):
    """Da a las columnas RECIÉN INSERTADAS el estilo de una columna vecina que
    ya venía en el fichero (por defecto la B), para no romper la banda cebra.

    `motor.insertar_columna()` desplaza lo que había; la columna nueva nace sin
    relleno, así que sin esto las filas pares del bloque quedan con un hueco
    blanco en mitad de la tabla.
    """
    for col in columnas:
        for fila in range(r0, r1 + 1):
            src = ws.cell(row=fila, column=modelo)
            ws.cell(row=fila, column=col)._style = copy.copy(src._style)


def _fmt(ws, ref, formato):
    for fila in ws[ref]:
        for cel in (fila if isinstance(fila, tuple) else (fila,)):
            cel.number_format = formato


def _limpiar_valores(ws, desde, hasta=None):
    """Vacía el CONTENIDO (no el estilo) de un rango de filas. Se usa en las
    hojas que se reconstruyen enteras."""
    hasta = hasta or ws.max_row
    for fila in range(desde, hasta + 1):
        for col in range(1, max(1, ws.max_column) + 1):
            ws.cell(row=fila, column=col).value = None


# ==========================================================================
# 06 — FIFO y caducidades
# ==========================================================================
#: Layout FINAL de `Control FIFO` (A:T). La v1.1 tenía A:K y las columnas
#: nuevas se insertan en `pre()` en cuatro tandas (§4).
CAB_FIFO = [
    '#',                                # A
    'Producto',                         # B
    'Categoría',                        # C  (DV del motor, §1.1)
    'Lote / Ref.',                      # D
    'Fecha de entrada',                 # E
    'Tipo de fecha',                    # F  (DV Caducidad/Consumo preferente)
    'Fecha de caducidad',               # G
    'Fecha de apertura',                # H
    'Vida útil tras abrir (días)',      # I
    'Fecha límite efectiva',            # J  = MIN(caducidad, apertura+vida)
    'Días restantes',                   # K
    'Estado',                           # L  (semáforo de 4 estados)
    'Días en almacén',                  # M  (da uso a «Fecha de entrada»)
    'Cantidad',                         # N
    'Unidad',                           # O
    'Precio/ud (€)',                    # P
    'Valor en riesgo (€)',              # Q
    'Zona de almacén',                  # R  (DV desde «Mapa Almacén»)
    'Posición',                         # S
    'Acción',                           # T
]

#: Inserciones sobre el layout de la v1.1 (A:K), en ORDEN. Cada tupla es
#: (índice donde insertar, cuántas). 1+3+5 = 9 columnas nuevas → A:T.
INSERTS_FIFO = [(6, 1), (8, 3), (13, 5)]

F_LIMITE = '=IF($G{f}="","",IF($H{f}="",$G{f},IF($I{f}="",$G{f},MIN($G{f},$H{f}+$I{f}))))'
F_DIAS = '=IF($J{f}="","",$J{f}-TODAY())'
F_ESTADO = ('=IF($K{f}="","",IF($K{f}<0,IF($F{f}="Consumo preferente",'
            '"⚠ REVISAR (consumo preferente)","⛔ CADUCADO — RETIRAR"),'
            'IF($K{f}<=2,"🔴 URGENTE",IF($K{f}<=7,"🟡 PRÓXIMO","🟢 OK"))))')
F_ALMACEN = '=IF($E{f}="","",TODAY()-$E{f})'
F_RIESGO = '=IF(OR($N{f}="",$P{f}=""),"",$N{f}*$P{f})'

TIPOS_FECHA = ['Caducidad', 'Consumo preferente']

#: RD-07/RT-21 · `Control FIFO` era la ÚNICA hoja nuclear del kit que se
#: entregaba COMPLETAMENTE vacía: 50 filas, 262 fórmulas nuevas y ni un solo
#: lote. El fichero que da nombre comercial al producto —«alertas de
#: caducidad»— abría con los cinco contadores a cero, el semáforo de cinco
#: estados sin demostrar y 20 columnas en blanco que el usuario tenía que
#: interpretar (¿qué es «Vida útil tras abrir»? ¿qué pongo en «Fecha límite
#: efectiva», que es calculada?). Todos los demás registros del kit sí traen
#: ejemplo: 04 cuatro líneas, 05 tres, 03 tres, 02 seis, 07 ocho, BONUS-09 ocho.
#:
#: Las fechas son RELATIVAS al día de generación (constantes en el fichero, no
#: fórmulas: las columnas son verdes y `TODAY()` ahí las volvería calculadas y
#: el cliente no podría escribir las suyas). Los seis lotes cubren los CINCO
#: estados y, de propina, el caso que explica la columna de apertura: un queso
#: con caducidad lejana cuya fecha límite efectiva la manda la vida útil tras
#: abrir.
#: (producto, categoría, lote, entrada (días atrás), tipo de fecha,
#:  caducidad (días desde hoy), apertura (días atrás o None), vida tras abrir,
#:  cantidad, unidad, precio/ud, zona, posición, acción)
LOTES_FIFO = [
    ('Merluza fresca', 'Pescados', 'L-2609-014', 4, 'Caducidad', -1,
     None, None, 4, 'kg', 14.00, 'Cámara pescado', 'Balda 1',
     'Retirar y anotar en el registro de mermas'),
    ('Nata 35 % M.G.', 'Lácteos', 'L-2608-221', 9, 'Consumo preferente', -2,
     None, None, 3, 'L', 3.10, 'Cámara elaborados', 'Balda 2',
     'Revisión organoléptica antes de decidir'),
    ('Pechuga de pollo', 'Cárnicos', 'L-2609-107', 1, 'Caducidad', 1,
     None, None, 6, 'kg', 6.50, 'Cámara crudos', 'Balda 1',
     'Al servicio de hoy'),
    ('Tomate', 'Verduras/Frutas', 'L-2609-088', 2, 'Caducidad', 5,
     None, None, 10, 'kg', 1.95, 'Cámara verduras', 'Cajón 3',
     'Priorizar en la producción de la semana'),
    ('Queso parmesano', 'Lácteos', 'L-2605-330', 7, 'Caducidad', 60,
     5, 10, 2, 'kg', 16.50, 'Cámara elaborados', 'Balda 3',
     'Abierto: manda la vida útil tras abrir, no la caducidad del envase'),
    ('Arroz redondo', 'Secos/Granos', 'L-2604-012', 40, 'Consumo preferente',
     300, None, None, 20, 'kg', 1.45, 'Economato seco', 'Estantería 2',
     'Sin riesgo: colocar detrás de lo que caduca antes'),
]

#: §4 — las 9 zonas del mapa de almacén, separando crudo de elaborados
#: (DOM-28). El nombre corto de la columna A alimenta la DV de «Zona de
#: almacén» del `Control FIFO`, así que no puede crecer: los 9 juntos ocupan
#: 118 de los 255 caracteres que Excel admite en una DV inline.
ZONAS = [
    ('Cámara crudos', '0-4 °C',
     'Carne, aves y caza crudas; siempre por debajo de todo lo demás',
     '2 estanterías',
     'Nunca en la misma cámara que producto listo para consumo.'),
    ('Cámara pescado', '0-2 °C',
     'Pescados y mariscos sobre hielo fundente', '1 mesa fría',
     'Hielo fundente con desagüe; reponer en cada servicio.'),
    ('Cámara elaborados', '0-4 °C',
     'Lácteos, elaborados propios, 5.ª gama y postres', '3 estanterías',
     'Todo tapado, etiquetado y fechado. Nunca junto a crudos.'),
    ('Cámara verduras', '4-8 °C',
     'Verduras y frutas', '2 estanterías',
     'Sin lavar hasta el momento de uso; aparta la fruta que emite etileno.'),
    ('Congelador', '≤ -18 °C',
     'Congelados', '1 arcón',
     'Anota la fecha de congelación. Nunca recongelar lo descongelado.'),
    ('Descongelación', '0-4 °C',
     'Producto en descongelación, identificado', '1 estantería',
     'Bandeja con rejilla, tapado y etiquetado EN DESCONGELACIÓN con fecha.'),
    ('Economato seco', 'Lugar fresco y seco, < 25 °C y HR < 60 %',
     'Secos y granos, conservas y bebidas no alcohólicas',
     '4 estanterías',
     'Producto a ≥ 10 cm del suelo y separado de la pared.'),
    ('Bodega', '12-16 °C',
     'Bebidas alcohólicas', '1 botellero',
     'Sin luz directa ni vibraciones; botellas tumbadas.'),
    ('Residuos', 'Ambiente, zona separada',
     'Residuos y envases', '3 contenedores',
     'Fuera del circuito de alimentos; cubos con tapa y pedal.'),
    # RD-04 · los huevos venían clasificados como «Lácteos», y el mapa manda
    # los lácteos a la cámara de elaborados (0-4 °C): el kit terminaba
    # diciéndole al usuario que refrigerara los huevos, justo lo contrario de
    # lo que dice su propia tabla legal de recepción dos ficheros más allá.
    # Refrigerar y volver a sacar el huevo provoca condensación en la cáscara
    # y es lo que el Reg. (CE) 589/2008 prohíbe antes de la venta.
    ('Huevera', 'Ambiente constante, 12-20 °C',
     'Huevos: NO refrigerar antes de la venta', '1 estantería',
     'Sin cambios bruscos de temperatura: la condensación en la cáscara '
     'facilita la entrada de Salmonella (Reg. (CE) 589/2008, art. 2).'),
    # RD-05 · el mapa guardaba los productos de limpieza y los químicos en el
    # mismo economato que las conservas y los granos. Separar químicos de
    # alimentos es de las primeras cosas que mira una inspección, y el propio
    # fichero ya lo sabía: para los residuos sí escribía «fuera del circuito
    # de alimentos». No había ninguna zona donde ponerlos.
    ('Almacén de químicos', 'Ambiente, zona separada y señalizada',
     'Productos de limpieza, desinfectantes y sus fichas de datos de '
     'seguridad', '1 armario',
     'Separado y señalizado; NUNCA sobre ni junto a alimentos ni sobre '
     'superficies de trabajo (Reg. (CE) 852/2004, Anexo II, Cap. IX).'),
]

#: §4 — los cinco estados del semáforo con su protocolo. El primero es el que
#: la v1.1 no tenía y por el que el fichero autorizaba a servir caducado.
ESTADOS_FIFO = [
    ('⛔ CADUCADO — RETIRAR', 'CADUCADO',
     'Retirada obligatoria y registro de merma en la plantilla 05. '
     'No se usa bajo ningún concepto.'),
    ('⚠ REVISAR (consumo preferente)', 'REVISAR',
     'Consumo preferente vencido: revisión organoléptica (aspecto, olor, '
     'textura) antes de decidir. No es automáticamente un desecho.'),
    ('🔴 URGENTE', 'URGENTE',
     'Quedan 2 días o menos: al servicio de hoy.'),
    ('🟡 PRÓXIMO', 'PRÓXIMO',
     'Quedan 7 días o menos: priorízalo en la producción de la semana.'),
    ('🟢 OK', 'OK',
     'Fuera de riesgo. Colócalo detrás de lo que caduca antes.'),
]

INSTRUCCIONES_06 = [
    'FIFO y Caducidades',
    'AI Chef Pro — aichef.pro',
    None,
    'Controla la rotación del género y el riesgo de caducidad, lote a lote.',
    None,
    'PESTAÑAS:',
    '1. Control FIFO — un lote por línea: entrada, caducidad, apertura, '
    'cantidad y valor en riesgo',
    '2. Alertas Caducidad — contadores por estado, valor en riesgo y '
    'protocolo de actuación',
    '3. Mapa Almacén — las 11 zonas con su temperatura; alimenta el '
    'desplegable de zona. Los químicos y la limpieza tienen su propio '
    'almacén, separado de los alimentos, y los huevos su propia zona a '
    'temperatura ambiente constante: refrigerarlos y volver a sacarlos '
    'provoca condensación en la cáscara',
    None,
    'FIFO Y FEFO NO SON LO MISMO:',
    'FIFO (First In First Out) es el criterio de COLOCACIÓN: lo que entra '
    'primero se coloca delante.',
    'FEFO (First Expired First Out) es el criterio de SALIDA, y es el que '
    'manda en cocina: sale primero lo que antes caduca.',
    'Esta plantilla calcula el estado por FEFO, sobre la fecha límite '
    'efectiva. La columna de días en almacén delata el lote olvidado al '
    'fondo de la cámara.',
    None,
    'CÓMO SE RELLENA CADA LÍNEA:',
    'Categoría, unidad y zona salen de desplegable: son las mismas 10 '
    'categorías en las 9 plantillas del kit y las mismas unidades en las 8 '
    'de producto. La de recepción (04) añade además la FAMILIA normativa, '
    'porque el límite legal de temperatura lo fija la norma y no tu '
    'categoría de compra: allí tienes el puente entre las dos listas.',
    'Tipo de fecha: Caducidad para los perecederos (carne, pescado, lácteos '
    'frescos, elaborados) y Consumo preferente para secos, conservas y '
    'congelados. Reglamento (UE) 1169/2011, artículo 24.',
    'Fecha de apertura y vida útil tras abrir: sólo para el envase ya '
    'abierto. La fecha límite efectiva se queda con la más cercana de las '
    'dos, que es la que manda.',
    'Cantidad y precio por unidad dan el valor en riesgo: es el número que '
    'hace que alguien actúe. Ojo a la hoja de alertas: una fila da el valor '
    'REALMENTE en riesgo (todo menos lo que está OK) y otra el valor total '
    'del stock registrado. No son lo mismo.',
    'Las seis primeras líneas vienen con lotes de EJEMPLO para que veas los '
    'cinco estados y las alertas funcionando desde el primer minuto. '
    'Bórralas antes de empezar con tu género.',
    None,
    'EL SEMÁFORO TIENE CINCO ESTADOS, NO TRES:',
    '⛔ CADUCADO — RETIRAR: fecha de caducidad vencida. Retirada obligatoria '
    'y registro de merma. No se usa bajo ningún concepto.',
    '⚠ REVISAR (consumo preferente): fecha de consumo preferente vencida. '
    'Revisión organoléptica antes de decidir.',
    '🔴 URGENTE: quedan 2 días o menos. Al servicio de hoy.',
    '🟡 PRÓXIMO: quedan 7 días o menos. Priorizar en la producción de la '
    'semana.',
    '🟢 OK: fuera de riesgo.',
    None,
    'El autofiltro de Control FIFO ya está activado: pincha la flecha de la '
    'columna Estado para ver sólo lo que te interesa.',
    None,
    motor.VERSION_LINE,
]


def _pre_06(wb, cambios):
    """Inserta las 9 columnas nuevas de `Control FIFO` (§4). Va en `pre()`
    porque el motor fija sus rangos —DV, formato condicional, verde— sobre el
    layout que encuentre."""
    ws = wb['Control FIFO']
    if str(ws['F4'].value or '').startswith('Tipo'):
        return                                          # ya insertadas
    for idx, cuantas in INSERTS_FIFO:
        for _ in range(cuantas):
            motor.insertar_columna(ws, idx)
    cambios.append('06:Control FIFO!A4:T4: +9 columnas (tipo de fecha, '
                   'apertura, vida útil tras abrir, fecha límite efectiva, '
                   'días en almacén, cantidad, unidad, precio/ud y valor en '
                   'riesgo) — DOM-25/DOM-26/TEC-12')


def _post_06(wb, cambios):
    _control_fifo_06(wb, cambios)
    _mapa_almacen_06(wb, cambios)
    _alertas_06(wb, cambios)
    _instrucciones(wb['Instrucciones'], INSTRUCCIONES_06)
    cambios.append('06:Instrucciones!A1:A31: reescritas — FIFO vs FEFO '
                   '(DOM-24/COM-26), los cuatro estados del semáforo '
                   '(DOM-13) y el Mapa de Almacén citado con el nombre real '
                   'de la pestaña, «Mapa Almacén» (§1.6)')


def _control_fifo_06(wb, cambios):
    ws = wb['Control FIFO']
    r0, r1 = 5, 54
    _cabecera(ws, 4, CAB_FIFO)
    _clonar_estilo(ws, [6, 8, 9, 10, 13, 14, 15, 16, 17], r0, r1)
    _nota(ws, 'A2', 'Una línea por LOTE, no por producto: dos entregas del '
                    'mismo género son dos líneas. Las SEIS primeras son un '
                    'ejemplo y cubren los cinco estados del semáforo (la '
                    'quinta enseña por qué existe la fecha de apertura): '
                    'bórralas antes de empezar. Fecha de revisión: '
                    '___/___/______   ·   Responsable: ________________')
    for f in range(r0, r1 + 1):
        _f(ws, 'J{}'.format(f), F_LIMITE.format(f=f))
        _f(ws, 'K{}'.format(f), F_DIAS.format(f=f))
        _f(ws, 'L{}'.format(f), F_ESTADO.format(f=f))
        _f(ws, 'M{}'.format(f), F_ALMACEN.format(f=f))
        _f(ws, 'Q{}'.format(f), F_RIESGO.format(f=f))
    _fmt(ws, 'N{}:N{}'.format(r0, r1), motor.FMT_CANT)
    _sembrar_lotes_06(ws)

    _limpiar_dv_c(ws)
    _dv(ws, 'F{}:F{}'.format(r0, r1), TIPOS_FECHA, 'Tipo de fecha no válido',
        'Caducidad = perecederos (carne, pescado, lácteos frescos). Consumo '
        'preferente = secos, conservas y congelados. Vencido no significa lo '
        'mismo en los dos casos.')
    _dv(ws, 'O{}:O{}'.format(r0, r1), motor.UNIDADES, 'Unidad no válida',
        'La misma unidad de compra que en las otras 8 plantillas del kit.')
    _dv(ws, 'R{}:R{}'.format(r0, r1), [z[0] for z in ZONAS],
        'Zona no válida',
        'Las {} zonas de la pestaña Mapa Almacén. Si cambias los nombres '
        'allí, cambia también este desplegable.'.format(len(ZONAS)))

    ws.auto_filter.ref = 'A4:T{}'.format(r1)
    _ancho(ws, [5, 26, 20, 14, 15, 18, 16, 16, 14, 16, 12, 30, 13, 11, 10,
                13, 17, 18, 12, 26])
    cambios.append(
        '06:Control FIFO!J5:Q54: 250 fórmulas nuevas — fecha límite efectiva '
        '(J), días restantes (K), semáforo de CUATRO estados con ⛔ CADUCADO '
        '— RETIRAR (L, DOM-13/TEC-11/COM-09), días en almacén (M, DOM-24) y '
        'valor en riesgo (Q, DOM-25); DV de tipo de fecha (DOM-26), unidad y '
        'zona; autofiltro en A4:T54 (DOM-27), que no existía en ninguna de '
        'las 30 hojas del kit')


def _sembrar_lotes_06(ws):
    """RD-07/RT-21 — seis lotes de ejemplo que dejan ver los CINCO estados a
    la vez y hacen que 'Alertas Caducidad' arranque con números."""
    hoy = datetime.date.today()
    for i, lote in enumerate(LOTES_FIFO):
        f = 5 + i
        (producto, categoria, ref, entrada, tipo, caducidad, apertura,
         vida, cantidad, unidad, precio, zona, posicion, accion) = lote
        ws.cell(row=f, column=2).value = producto + ' (ejemplo)'
        ws.cell(row=f, column=3).value = categoria
        ws.cell(row=f, column=4).value = ref
        ws.cell(row=f, column=5).value = hoy - datetime.timedelta(entrada)
        ws.cell(row=f, column=6).value = tipo
        ws.cell(row=f, column=7).value = hoy + datetime.timedelta(caducidad)
        if apertura is not None:
            ws.cell(row=f, column=8).value = (hoy
                                              - datetime.timedelta(apertura))
            ws.cell(row=f, column=9).value = vida
        ws.cell(row=f, column=14).value = cantidad
        ws.cell(row=f, column=15).value = unidad
        ws.cell(row=f, column=16).value = precio
        ws.cell(row=f, column=18).value = zona
        ws.cell(row=f, column=19).value = posicion
        ws.cell(row=f, column=20).value = accion
        for col in (5, 7, 8):
            ws.cell(row=f, column=col).number_format = motor.FMT_FECHA
        ws.cell(row=f, column=16).number_format = motor.FMT_EUR


def _mapa_almacen_06(wb, cambios):
    ws = wb['Mapa Almacén']
    ws['A1'] = 'MAPA DE ALMACÉN — ZONAS DE ALMACENAMIENTO'
    _cabecera(ws, 3, ['Zona', 'Temperatura', 'Qué se guarda aquí',
                      'Capacidad', 'Notas'])
    # La v1.1 llegaba hasta la fila 9: las filas nuevas heredan el estilo de
    # dos filas más arriba para que la banda cebra siga alternando (copiar
    # siempre de la 9 dejaría varias filas del mismo color).
    for f in range(10, 4 + len(ZONAS)):
        for c in range(1, 6):
            modelo = ws.cell(row=f - 2, column=c)
            ws.cell(row=f, column=c)._style = copy.copy(modelo._style)
    for i, zona in enumerate(ZONAS):
        f = 4 + i
        for c, valor in enumerate(zona):
            ws.cell(row=f, column=1 + c, value=valor)
    _nota(ws, 'A{}'.format(4 + len(ZONAS) + 1),
          'Zonas de ejemplo: cámbialas por las tuyas. Los nombres de la '
          'columna Zona son los que salen en el desplegable de Control FIFO, '
          'así que si los editas, edita también ese desplegable. La '
          'separación crudo / elaborado no es opcional, y los productos de '
          'limpieza y los químicos NUNCA van en el economato de los '
          'alimentos: son las dos primeras cosas que mira una inspección.')
    _ancho(ws, [22, 34, 46, 16, 52])
    cambios.append(
        '06:Mapa Almacén!A4:E{}: {} zonas precargadas separando crudo de '
        'elaborados (DOM-28/COM-15) — la hoja se vendía tres veces en la '
        'landing y se entregaba con seis filas EN BLANCO; alimenta la DV de '
        'zona del Control FIFO. Ronda 2: se separan los QUÍMICOS del '
        'economato (RD-05) y los huevos estrenan zona de ambiente estable, '
        'porque el mapa mandaba refrigerarlos (RD-04)'
        .format(3 + len(ZONAS), len(ZONAS)))


def _alertas_06(wb, cambios):
    """DOM-27 — la hoja pasa de seis líneas de texto a contadores reales."""
    ws = wb['Alertas Caducidad']
    _limpiar_valores(ws, 1)
    ws['A1'] = '⚠️ ALERTAS DE CADUCIDAD'
    ws['A1'].font = Font(bold=True, size=13, color='9C0006')
    _nota(ws, 'A2', 'Todo se calcula solo desde Control FIFO. Aquí no se '
                    'escribe nada.')
    _cabecera(ws, 3, ['Estado', 'Nº de lotes', 'Valor en riesgo (€)',
                      'Qué hacer'])
    for i, dato in enumerate(ESTADOS_FIFO):
        etiqueta, clave, protocolo = dato
        f = 4 + i
        ws.cell(row=f, column=1, value=etiqueta)
        _f(ws, 'B{}'.format(f),
           '=COUNTIF(\'Control FIFO\'!$L$5:$L$54,"*{}*")'.format(clave))
        _f(ws, 'C{}'.format(f),
           '=SUMIF(\'Control FIFO\'!$L$5:$L$54,"*{}*",'
           '\'Control FIFO\'!$Q$5:$Q$54)'.format(clave))
        cel = ws.cell(row=f, column=4, value=protocolo)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
    # RD-14 · la fila de totales presentaba el valor de TODO el stock
    # registrado bajo la cabecera «Valor en riesgo (€)». En un almacén con
    # 4.000 € de género y 200 € realmente en riesgo, la hoja de alertas
    # remataba diciendo «valor en riesgo: 4.000 €»: un número así en la hoja
    # que se enseña al propietario o al inspector es peor que no darlo,
    # porque las cinco filas de arriba sí están bien calculadas por estado y
    # ésta las desmentía. Ahora son DOS filas y cada una dice lo que es.
    ws['A9'] = 'VALOR REALMENTE EN RIESGO (todo menos 🟢 OK)'
    ws['A9'].font = Font(bold=True, color='9C0006')
    _f(ws, 'B9', '=SUM($B$4:$B$7)')
    _f(ws, 'C9', '=SUM($C$4:$C$7)')
    ws['D9'] = ('Es la suma de los cuatro estados de arriba, sin los que '
                'están fuera de riesgo. Éste es el número que mueve a '
                'alguien.')
    ws['A10'] = 'TOTAL de lotes registrados / valor total del stock'
    ws['A10'].font = Font(bold=True)
    # COUNTIF sobre la columna PRODUCTO, no sobre el estado: el estado es una
    # fórmula que devuelve "" en las filas vacías y COUNTIF(…,"<>") las
    # contaría como llenas, dando 50 siempre (§1.1, sustituto de COUNTA).
    _f(ws, 'B10', '=COUNTIF(\'Control FIFO\'!$B$5:$B$54,"<>")')
    _f(ws, 'C10', '=SUM(\'Control FIFO\'!$Q$5:$Q$54)')
    ws['D10'] = ('Todo el stock que tienes registrado en Control FIFO, esté '
                 'o no en riesgo. No lo confundas con la fila de arriba.')
    for fila in (9, 10):
        for col in ('B', 'C'):
            ws['{}{}'.format(col, fila)].font = Font(bold=True)
        ws['D{}'.format(fila)].alignment = Alignment(wrap_text=True,
                                                     vertical='top')
    _fmt(ws, 'B4:B10', motor.FMT_ENT)
    _fmt(ws, 'C4:C10', motor.FMT_EUR)

    _titulo(ws, 'A12', 'PROTOCOLO DE ACTUACIÓN')
    protocolo = [
        '1. ⛔ CADUCADO — RETIRAR: retirada obligatoria del producto, baja '
        'del stock y registro en la plantilla 05 (Control de Mermas). Un '
        'producto con la fecha de caducidad vencida no se usa bajo ningún '
        'concepto, ni para personal ni para elaboraciones.',
        '2. ⚠ REVISAR (consumo preferente): la fecha de consumo preferente '
        'vencida NO obliga a tirar. Revisión organoléptica (aspecto, olor, '
        'textura) y decisión del responsable; si hay duda, se retira.',
        '3. 🔴 URGENTE: al servicio de HOY. Anótalo en el briefing y '
        'colócalo delante de todo.',
        '4. 🟡 PRÓXIMO: priorízalo en la producción de la semana y revísalo '
        'al inicio de cada turno de mañana.',
    ]
    for i, linea in enumerate(protocolo):
        cel = ws.cell(row=13 + i, column=1, value=linea)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
    _nota(ws, 'A18',
          'Para verlos uno a uno: en Control FIFO, flecha del encabezado '
          'Estado y filtra por el que quieras. El autofiltro ya está puesto.')
    _ancho(ws, [46, 14, 20, 70])
    cambios.append(
        '06:Alertas Caducidad!A3:D10 + A12:A18: contadores COUNTIF por '
        'estado y valor en riesgo con SUMIF, más el protocolo en cuatro '
        'líneas con caducado = retirada obligatoria (DOM-27/COM-09) — la hoja '
        'era una instrucción de trabajo manual con 0 fórmulas. Ronda 2 '
        '(RD-14): el «valor en riesgo» del pie era el valor de TODO el stock; '
        'ahora hay una fila para cada cosa')


# ==========================================================================
# 07 — análisis de costes de compras
# ==========================================================================
#: Layout FINAL de `Top 20 Productos` (A:K) + bloque «Top 5» en L:M.
CAB_TOP20 = [
    '#',                       # A
    'Producto',                # B
    'Categoría',               # C  (DV del motor, §1.1)
    'Unidad',                  # D  (DV de este grupo)
    'Gasto mensual (€)',       # E
    # RD-27 · un jefe de compras no firma «gasto anual» a partir de un mes
    # en un negocio estacional (el solomillo de enero pasaba a 11.520 € al
    # año sin decir de dónde salía). La cabecera lo dice ahora.
    'Gasto anual estimado (€) — mes × 12',   # F  = E × 12
    '% del Top 20',            # G  = E / E24
    'Precio anterior (€)',     # H
    'Precio actual (€)',       # I
    'Variación de precio (%)',  # J  = I/H − 1  (CF > 5 %, §1.2). El «(%)»
                                #    no es cosmético: sin él, «precio» es
                                #    clave de MONEDA en el motor y la
                                #    variación se imprimiría «0,05 €».

    'Proveedor',               # K
]

#: (índice donde insertar) sobre el layout de la v1.1 (A:I), en ORDEN.
INSERTS_TOP20 = [4, 8]

#: §4 — 8 filas de ejemplo, coherentes producto a producto con las 50 del 01
#: (mismos nombres, mismas unidades, mismos precios de referencia) y con los
#: 6 proveedores del 02. Ordenadas de mayor a menor gasto, que es lo que la
#: hoja promete.
#:
#: Ronda 2 · RD-21 — los consumos de la calculadora del BONUS-09 no cuadraban
#: con este gasto mensual para los mismos productos y los mismos precios, con
#: desviaciones de hasta el 50 %: un cliente que hace la regla de tres —y un
#: jefe de compras la hace— encontraba el descuadre en el primer minuto. Ahora
#: el gasto NO se teclea: sale de `motor.CONSUMO_EJEMPLO`, la misma tabla de
#: la que salen los par levels del 01 y los consumos del BONUS-09.
#:     gasto mensual = consumo diario × 30 × precio actual
#: (producto, categoría, unidad, precio anterior, precio actual)
_TOP20_BASE = [
    ('Solomillo de ternera', 'Cárnicos', 'kg', 30.50, 32.00),
    ('Salmón fresco', 'Pescados', 'kg', 12.90, 14.00),
    ('Gambas', 'Pescados', 'kg', 17.50, 18.00),
    ('Pechuga de pollo', 'Cárnicos', 'kg', 6.20, 6.50),
    ('Cerveza de grifo (barril 30 L)', 'Bebidas Alcohólicas', 'barril',
     53.00, 55.00),
    ('Aceite de oliva virgen extra', 'Secos/Granos', 'L', 5.10, 5.80),
    ('Queso parmesano', 'Lácteos', 'kg', 16.20, 16.50),
    ('Patata', 'Verduras/Frutas', 'kg', 0.95, 0.85),
]

#: (producto, categoría, unidad, gasto mensual, precio anterior, precio actual)
TOP20_EJEMPLO = [
    (nombre, categoria, unidad,
     round(motor.CONSUMO_EJEMPLO[nombre][0] * 30 * ahora, 2), antes, ahora)
    for nombre, categoria, unidad, antes, ahora in _TOP20_BASE
]

#: RC-07 · con el sufijo «(ejemplo)», que es como aparecen en el directorio
#: del 02 y en el desplegable del 03: si uno escribe «Cárnicas del Norte» y el
#: desplegable ofrece «Cárnicas del Norte (ejemplo)», filtrar no agrupa.
PROVEEDOR_TOP20 = {
    'Cárnicos': motor.PROVEEDORES_MARCADOS[0],
    'Pescados': motor.PROVEEDORES_MARCADOS[1],
    'Verduras/Frutas': motor.PROVEEDORES_MARCADOS[2],
    'Secos/Granos': motor.PROVEEDORES_MARCADOS[3],
    'Lácteos': motor.PROVEEDORES_MARCADOS[3],
    'Bebidas Alcohólicas': motor.PROVEEDORES_MARCADOS[4],
}

#: §4 — compras de ejemplo de UN mes (enero) por las 10 categorías canónicas.
#: Suman 10.400 €, de los que el Top 20 de ejemplo se lleva 6.474,50 € (62 %),
#: que es el reparto de Pareto habitual en un economato de restaurante.
COMPRAS_ENE = [2850.00, 1980.00, 620.00, 1150.00, 1240.00, 480.00, 980.00,
               540.00, 380.00, 180.00]

MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
         'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

#: `Dashboard KPIs` — (fila, etiqueta, objetivo, cómo se calcula).
#: Las fórmulas van aparte porque cada una es distinta.
#: RT-19 · los objetivos que la v2.0 dejó sin parametrizar eran literales de
#: TEXTO («28-32 %», «< ±5 %») en celdas BLOQUEADAS dentro de una hoja
#: protegida: el cliente no podía cambiar su propio objetivo sin desproteger,
#: y encima no eran computables, así que no había semáforo posible. Ahora son
#: NÚMEROS en celda verde, que es la doctrina del kit («parámetros en celda,
#: nunca literales en la fórmula»).
#: (fila, etiqueta, objetivo, formato, editable, cómo se calcula)
KPIS = [
    (4, 'Food cost (%)', 0.32, motor.FMT_PCT1, True,
     '(Existencias iniciales + Compras del periodo − Existencias finales) / '
     'Ventas. Es CONSUMO, no compras. El rango sano de un restaurante está '
     'entre el 28 % y el 32 %: aquí se fija el TECHO, cámbialo por el tuyo.'),
    (5, 'Coste por cubierto (€)', None, motor.FMT_EUR, False,
     'Consumo de materia prima (sin Limpieza ni Otros) / cubiertos '
     'servidos. El objetivo es el 30 % de tu ticket medio, no una cifra '
     'fija en euros.'),
    (6, 'Compras del periodo (€)', '—', None, False,
     'Total del mes que elijas arriba en la pestaña Coste por Categoría.'),
    (7, 'Variación vs periodo anterior', 0.05, motor.FMT_PCT1, True,
     'Compras del periodo / compras del periodo anterior − 1. Se compara en '
     'valor absoluto: una caída del 20 % también hay que explicarla.'),
    (8, 'Productos con subida > 5 %', 0, motor.FMT_ENT, True,
     'Líneas del Top 20 cuya variación de precio supera el 5 %.'),
    (9, 'Mayor subida de precio', '—', None, False,
     'Producto del Top 20 con la variación más alta.'),
]

#: RD-27 · el desplegable del periodo. El KPI de compras leía SIEMPRE el total
#: de los doce meses, así que el dashboard no podía enseñar nunca un mes
#: concreto salvo que sólo hubiera uno cargado — y la vista mensual, la que se
#: mira cada día 1, era imposible.
PERIODOS = ['Año completo'] + ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                               'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

#: Entradas verdes `B12:B17` (§4). Los valores de ejemplo son coherentes con
#: las compras de enero de arriba: 10.400 € de compras, 34.000 € de ventas y
#: 1.450 cubiertos dan un food cost del 31,0 % y un ticket medio de 23,45 €.
ENTRADAS_KPI = [
    (12, 'Ventas del periodo (€, sin IVA)', 34000.00, motor.FMT_EUR,
     'Base imponible, sin IVA. Mismo periodo que las compras cargadas.'),
    (13, 'Cubiertos servidos', 1450, motor.FMT_ENT,
     'Comensales reales del periodo, no reservas.'),
    (14, 'Existencias iniciales (€)', 4200.00, motor.FMT_EUR,
     'Valor del inventario al empezar el periodo (plantilla 01 o BONUS-08).'),
    (15, 'Existencias finales (€)', 4050.00, motor.FMT_EUR,
     'Valor del inventario al cerrar el periodo.'),
    (16, 'Compras del periodo anterior (€)', 9980.00, motor.FMT_EUR,
     'Para la variación. Sin IVA, igual que el resto.'),
    (17, 'Ticket medio (€)', 23.45, motor.FMT_EUR,
     'Ventas / cubiertos. Fija el objetivo de coste por cubierto.'),
]

INSTRUCCIONES_07 = [
    'Análisis de Costes de Compras',
    'AI Chef Pro — aichef.pro',
    None,
    'Mira en qué se te va el dinero de las compras y calcula tus KPIs de '
    'coste.',
    None,
    'TODOS LOS IMPORTES VAN SIN IVA (BASE IMPONIBLE).',
    'El IVA soportado se deduce: no forma parte del food cost. Si copias el '
    'total de la factura en vez de la base imponible, tu food cost sale '
    'entre un 4 % y un 21 % inflado y todas las decisiones que tomes con él '
    'serán erróneas.',
    None,
    'PESTAÑAS:',
    '1. Coste por Categoría — las 10 categorías del kit, mes a mes; es la '
    'tabla que alimenta a todas las demás',
    '2. Evolución Mensual — el total de cada mes, su variación y el '
    'acumulado del año',
    '3. Top 20 Productos — los productos que más dinero se llevan, con la '
    'alerta de subida de precio y el ranking Top 5, que se ordena solo (no '
    'hace falta que ordenes la tabla)',
    '4. Dashboard KPIs — food cost, coste por cubierto y variación, '
    'calculados de verdad',
    None,
    'CÓMO SE USA:',
    'Carga cada mes el gasto por categoría en Coste por Categoría. El resto '
    'del libro se calcula solo.',
    'En el Dashboard eliges el PERIODO en una casilla verde: un mes '
    'concreto o «Año completo». Las ventas, los cubiertos y las existencias '
    'que pongas tienen que ser las de ese mismo periodo, y los objetivos '
    '(food cost, variación, subidas de precio) también son tuyos: son '
    'casillas verdes, cámbialos.',
    'La columna Estado del Dashboard es un semáforo de verdad: verde dentro '
    'del objetivo, ámbar hasta un 10 % por encima y rojo a partir de ahí.',
    'El food cost real es de CONSUMO, no de compras: existencias iniciales '
    'más compras menos existencias finales. Por eso el dashboard pide las '
    'dos existencias, que salen de la plantilla 01 o del BONUS-08.',
    'El coste por cubierto deja fuera Limpieza y Otros: no son materia '
    'prima. Y su objetivo no es una cifra fija en euros, sino un porcentaje '
    'de tu ticket medio (el 30 % que trae por defecto).',
    None,
    'Las filas de ejemplo llevan datos inventados para que veas el libro '
    'funcionando: bórralos y pon los tuyos.',
    None,
    motor.VERSION_LINE,
]


def _pre_07(wb, cambios):
    ws = wb['Top 20 Productos']
    if str(ws['D3'].value or '').startswith('Unidad'):
        return
    for idx in INSERTS_TOP20:
        motor.insertar_columna(ws, idx)
    cambios.append('07:Top 20 Productos!A3:K3: +2 columnas (unidad de compra '
                   'y precio anterior) — sin precio anterior la columna '
                   '«Variación vs Anterior» era literalmente incalculable '
                   '(DOM-29)')


def _post_07(wb, cambios):
    _coste_categoria_07(wb, cambios)
    _evolucion_07(wb, cambios)
    _top20_07(wb, cambios)
    _dashboard_07(wb, cambios)
    _instrucciones(wb['Instrucciones'], INSTRUCCIONES_07)
    cambios.append('07:Instrucciones!A1:A24: reescritas — aviso de «sin IVA» '
                   '(DOM-30), definición correcta de los KPIs (COM-23) y la '
                   'pestaña Evolución Mensual, que se citaba desde la v1.1 y '
                   'no existía (§1.6)')


def _coste_categoria_07(wb, cambios):
    ws = wb['Coste por Categoría']
    ws['A1'] = 'COSTE DE COMPRAS POR CATEGORÍA (€, SIN IVA)'
    _nota(ws, 'A2', 'Todos los importes SIN IVA (base imponible). La columna '
                    'de enero trae datos de ejemplo: bórralos y carga los '
                    'tuyos.')
    ws['N3'] = 'Total año (€, sin IVA)'
    ws['O3'] = '% del Total'
    for i, importe in enumerate(COMPRAS_ENE):
        ws.cell(row=4 + i, column=2, value=importe)
    for f in range(4, 15):
        _f(ws, 'O{}'.format(f), '=IFERROR($N{f}/$N$14,"")'.format(f=f))
    ws['O14'].number_format = motor.FMT_PCT1
    ws['O14'].font = Font(bold=True)
    _nota(ws, 'A16',
          'La fila 14 (TOTAL) es la que lee el Dashboard KPIs. Allí eliges en '
          'una casilla verde QUÉ periodo mides: un mes concreto o el año '
          'completo. Las ventas, los cubiertos y las existencias que pongas '
          'en el dashboard tienen que ser las de ese mismo periodo.')
    cambios.append(
        '07:Coste por Categoría!O4:O14 + B4:B13: «% del Total» tenía el '
        'denominador calculado al lado (N14) y las 11 celdas vacías '
        '(DOM-31/TEC-08); cabecera y título con «(€, sin IVA)» (DOM-30) y '
        'enero sembrado con 10 importes de ejemplo')


def _evolucion_07(wb, cambios):
    """§1.6 — `Evolución Mensual` se citaba en `Instrucciones!A8` desde la v1.1
    y NO EXISTÍA. Es un registro simple derivado de la fila TOTAL."""
    nueva = 'Evolución Mensual' not in wb.sheetnames
    if nueva:
        ws = wb.create_sheet('Evolución Mensual', 2)
    else:
        ws = wb['Evolución Mensual']
    ws['A1'] = 'EVOLUCIÓN MENSUAL DEL GASTO EN COMPRAS (€, SIN IVA)'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=motor.CAB)
    if 'A1:D1' not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells('A1:D1')
    _nota(ws, 'A2', 'No se escribe nada aquí: sale entero de la fila TOTAL de '
                    'Coste por Categoría.')
    _cabecera(ws, 3, ['Mes', 'Compras del mes (€)',
                      'Variación vs mes anterior (%)',
                      'Acumulado del año (€)'])
    for i, mes in enumerate(MESES):
        f = 4 + i
        col = motor.get_column_letter(2 + i)          # B..M de la otra hoja
        ws.cell(row=f, column=1, value=mes)
        if f % 2 == 1:                              # banda cebra del kit
            for c in range(1, 5):
                ws.cell(row=f, column=c).fill = PatternFill(
                    'solid', fgColor=motor.BANDA)
        _f(ws, 'B{}'.format(f),
           "='Coste por Categoría'!{}$14".format(col))
        if i == 0:
            ws['C4'] = None
        else:
            _f(ws, 'C{}'.format(f),
               '=IF(OR($B{a}=0,$B{f}=0),"",IFERROR($B{f}/$B{a}-1,""))'
               .format(f=f, a=f - 1))
        _f(ws, 'D{}'.format(f), '=SUM($B$4:$B{})'.format(f))
    _nota(ws, 'A17', 'La variación se deja en blanco mientras alguno de los '
                     'dos meses esté a cero: un mes sin cargar no es una '
                     'caída del 100 %.')
    _ancho(ws, [18, 24, 30, 24])
    cambios.append(
        '07:Evolución Mensual: hoja CREADA con 12 meses (§1.6/TEC-05) — '
        'Instrucciones!A8 la citaba desde la v1.1 y no existía; B = total '
        'del mes, C = variación mes a mes con guarda de cero, D = acumulado')


#: RD-08 · el bloque «Top 5» va por DEBAJO de la fila TOTAL (24), fuera del
#: rango de datos que el usuario puede ordenar y fuera del bloque de verde.
FILA_TOP5 = 26


def _top20_07(wb, cambios):
    ws = wb['Top 20 Productos']
    r0, r1, total = 4, 23, 24
    _cabecera(ws, 3, CAB_TOP20)
    _clonar_estilo(ws, [4, 8], r0, r1)
    # RD-08/RT-11 · la hoja pedía por escrito «ordena la tabla de mayor a
    # menor gasto», y ordenarla era a la vez IMPOSIBLE (la hoja está protegida
    # y el rango contiene columnas bloqueadas: Excel se niega aunque el
    # permiso «Ordenar» esté concedido) y DESTRUCTIVO (el bloque «Top 5» vivía
    # dentro de las mismas filas 4-8 del rango de datos, con el k de LARGE
    # literal por fila, así que al ordenar se barajaba el propio ranking). La
    # única instrucción de uso de la hoja era la que la rompía.
    _nota(ws, 'A2', NOTA_EJEMPLO + ' NO hace falta que ordenes la tabla: el '
                                   'ranking Top 5 de abajo la ordena solo con '
                                   'LARGE, escribas en el orden que escribas.')
    for i, dato in enumerate(TOP20_EJEMPLO):
        f = r0 + i
        producto, categoria, unidad, gasto, antes, ahora = dato
        ws.cell(row=f, column=2, value=producto)
        ws.cell(row=f, column=3, value=categoria)
        ws.cell(row=f, column=4, value=unidad)
        ws.cell(row=f, column=5, value=gasto)
        ws.cell(row=f, column=8, value=antes)
        ws.cell(row=f, column=9, value=ahora)
        ws.cell(row=f, column=11,
                value=PROVEEDOR_TOP20.get(categoria) or None)
    for f in range(r0, r1 + 1):
        _f(ws, 'F{}'.format(f), '=IF($E{f}="","",$E{f}*12)'.format(f=f))
        _f(ws, 'G{}'.format(f), '=IFERROR($E{f}/$E${t},"")'
           .format(f=f, t=total))
        _f(ws, 'J{}'.format(f),
           '=IF(OR($H{f}="",$I{f}=""),"",$I{f}/$H{f}-1)'.format(f=f))
    ws['A{}'.format(total)] = 'TOTAL'
    ws['A{}'.format(total)].font = Font(bold=True)
    _f(ws, 'E{}'.format(total), '=SUM($E${a}:$E${b})'.format(a=r0, b=r1))
    _f(ws, 'F{}'.format(total), '=SUM($F${a}:$F${b})'.format(a=r0, b=r1))
    _f(ws, 'G{}'.format(total), '=IFERROR($E${t}/$E${t},"")'.format(t=total))
    for col in ('E', 'F'):
        ws['{}{}'.format(col, total)].number_format = motor.FMT_EUR
        ws['{}{}'.format(col, total)].font = Font(bold=True)
    ws['G{}'.format(total)].number_format = motor.FMT_PCT1
    ws['G{}'.format(total)].font = Font(bold=True)

    # Bloque «Top 5 por gasto» (§4). El `k` de LARGE va LITERAL por fila: es
    # lo que permite que las cinco fórmulas sean independientes y que pycel
    # las cachee. Con dos productos de gasto idéntico, MATCH devuelve siempre
    # el primero — está dicho en la nota de abajo.
    #
    # RD-08 · vivía en L4:M8, es decir DENTRO de las filas 4-8 del rango de
    # datos. Cualquiera que ordenara la tabla —lo que la propia hoja mandaba
    # hacer— barajaba el ranking. Ahora va debajo de la fila TOTAL, fuera del
    # rango ordenable y fuera del bloque que el motor pinta de verde.
    _cabecera(ws, FILA_TOP5, ['Top 5 productos', 'Gasto mensual (€)'],
              col0=12)
    for k in range(1, 6):
        f = FILA_TOP5 + k
        _f(ws, 'L{}'.format(f),
           '=IFERROR(INDEX($B${a}:$B${b},MATCH(LARGE($E${a}:$E${b},{k}),'
           '$E${a}:$E${b},0)),"")'.format(a=r0, b=r1, k=k))
        _f(ws, 'M{}'.format(f),
           '=IFERROR(LARGE($E${a}:$E${b},{k}),"")'.format(a=r0, b=r1, k=k))
        ws.cell(row=f, column=13).number_format = motor.FMT_EUR
    _nota(ws, 'A{}'.format(FILA_TOP5 + 7),
          'El ranking Top 5 de la derecha se calcula solo con LARGE: no hace '
          'falta que ordenes la tabla y, de hecho, la hoja está protegida '
          'para que no la desordenes sin querer. Si dos productos gastan '
          'exactamente lo mismo, enseña dos veces el primero; cambia un '
          'céntimo y se separan.')
    # Si el fichero venía de una pasada anterior con el Top 5 en L4:M8, esas
    # celdas tienen que quedar vacías o el bloque saldría duplicado.
    for f in range(3, 9):
        for c in (12, 13):
            ws.cell(row=f, column=c).value = None
    _limpiar_dv_c(ws)
    _dv(ws, 'D{}:D{}'.format(r0, r1), motor.UNIDADES, 'Unidad no válida',
        'La unidad en la que compras el producto: es la que dan los precios '
        'anterior y actual de las columnas H e I.')
    _ancho(ws, [5, 30, 22, 11, 17, 17, 13, 18, 17, 16, 30, 30, 18])
    cambios.append(
        '07:Top 20 Productos!B4:M24: la hoja tenía 0 fórmulas y sólo la '
        'numeración 1-20 (TEC-08/COM-11). Ahora es la tabla de entrada '
        '(§7-bis.1): gasto anual (F), % del Top 20 (G), variación de precio '
        'con CF > 5 % (J, DOM-29), 8 productos de ejemplo coherentes con el '
        '01 y ranking Top 5 con LARGE/INDEX/MATCH en L4:M8')


def _dashboard_07(wb, cambios):
    ws = wb['Dashboard KPIs']
    ws['A1'] = 'DASHBOARD — KPIs DE COMPRAS'
    _nota(ws, 'A2', 'Elige el periodo y rellena las casillas verdes. Todo lo '
                    'demás se calcula solo, con los importes SIN IVA. La '
                    'columna Estado te dice de un vistazo si cada KPI está '
                    'donde tiene que estar.')
    # «Valor Actual» contiene «valor», que es clave de MONEDA en
    # `motor.FORMATO_POR_CABECERA`: dejaría el food cost (un porcentaje) y el
    # nombre del producto de B9 con formato de euros.
    _cabecera(ws, 3, ['KPI', 'Resultado', 'Objetivo', 'Cómo se calcula',
                      'Estado'])
    for fila, etiqueta, objetivo, fmt, editable, como in KPIS:
        ws.cell(row=fila, column=1, value=etiqueta)
        if objetivo is not None:
            cel = ws.cell(row=fila, column=3, value=objetivo)
            if fmt:
                cel.number_format = fmt
            if editable:
                # RT-19 · objetivo del cliente, no del producto.
                motor.marcar_verde(ws, 'C{}'.format(fila))
        cel = ws.cell(row=fila, column=4, value=como)
        cel.alignment = Alignment(wrap_text=True, vertical='top')

    _f(ws, 'B4', '=IFERROR(($B$14+$B$6-$B$15)/$B$12,"")')
    _f(ws, 'B5', '=IFERROR(($B$14+$B$6-$B$15'
                 "-'Coste por Categoría'!$N$12"
                 "-'Coste por Categoría'!$N$13)/$B$13,\"\")")
    # RD-27 · el KPI de compras deja de leer SIEMPRE el total de los doce
    # meses: lee el mes elegido en B10, o el año entero si así se pide.
    _f(ws, 'B6', '=IF($B$10="","",IF($B$10="Año completo",'
                 "'Coste por Categoría'!$N$14,"
                 "IFERROR(INDEX('Coste por Categoría'!$B$14:$M$14,"
                 "MATCH($B$10,'Coste por Categoría'!$B$3:$M$3,0)),\"\")))")
    _f(ws, 'B7', '=IFERROR($B$6/$B$16-1,"")')
    _f(ws, 'B8', '=COUNTIF(\'Top 20 Productos\'!$J$4:$J$23,">0.05")')
    _f(ws, 'B9', "=IFERROR(INDEX('Top 20 Productos'!$B$4:$B$23,"
                 "MATCH(MAX('Top 20 Productos'!$J$4:$J$23),"
                 "'Top 20 Productos'!$J$4:$J$23,0)),\"\")")
    # DOM-31/COM-23 — el objetivo del coste por cubierto deja de ser «< 4,50 €»
    # (que sólo vale para un ticket de 15 €) y pasa a ser el 30 % del ticket.
    _f(ws, 'C5', '=IFERROR($B$17*0.3,"")')

    # ---- RD-13/RT-06/RC-13 · la columna de Estado que faltaba ------------
    # Era el ÚNICO dashboard del kit sin semáforo —justo el que la landing
    # vende como «dashboard de KPIs»— y encima su print_area ya reservaba una
    # columna E que nunca se llenó: imprimía una columna fantasma. La SPEC
    # §1.2 la exige por su nombre («07!'Dashboard KPIs'!E»).
    _f(ws, 'E4', '=IF(OR($B$4="",$C$4=""),"",IF($B$4<=$C$4,"🟢 OK",'
                 'IF($B$4<=$C$4*1.1,"🟡 REVISAR","🔴 ALERTA")))')
    _f(ws, 'E5', '=IF(OR($B$5="",$C$5=""),"",IF($B$5<=$C$5,"🟢 OK",'
                 'IF($B$5<=$C$5*1.1,"🟡 REVISAR","🔴 ALERTA")))')
    _f(ws, 'E7', '=IF(OR($B$7="",$C$7=""),"",IF(ABS($B$7)<=$C$7,"🟢 OK",'
                 'IF(ABS($B$7)<=$C$7*2,"🟡 REVISAR","🔴 ALERTA")))')
    _f(ws, 'E8', '=IF(OR($B$8="",$C$8=""),"",IF($B$8<=$C$8,"🟢 OK",'
                 'IF($B$8<=2,"🟡 REVISAR","🔴 ALERTA")))')
    for coord, fmt in (('B4', motor.FMT_PCT1), ('B5', motor.FMT_EUR),
                       ('B6', motor.FMT_EUR), ('B7', motor.FMT_PCT1),
                       ('B8', motor.FMT_ENT), ('C5', motor.FMT_EUR)):
        ws[coord].number_format = fmt
        ws[coord].font = Font(bold=True)
    for fila in range(4, 10):
        ws.cell(row=fila, column=5).font = Font(bold=True)
        ws.cell(row=fila, column=5).alignment = Alignment(
            horizontal='center', vertical='center')

    # ---- RD-27 · selector de periodo -------------------------------------
    ws['A10'] = 'PERIODO QUE SE MIDE'
    ws['A10'].font = Font(bold=True)
    ws['B10'] = 'Ene'
    motor.marcar_verde(ws, 'B10')
    ws['D10'] = ('Elige el mes que has cargado en Coste por Categoría, o '
                 '«Año completo». Las ventas, los cubiertos y las '
                 'existencias que pongas abajo tienen que ser las de ESE '
                 'periodo.')
    ws['D10'].alignment = Alignment(wrap_text=True, vertical='top')
    _limpiar_dv_c(ws)
    _dv(ws, 'B10', PERIODOS, 'Periodo no válido',
        'Año completo o uno de los doce meses de la pestaña Coste por '
        'Categoría.')

    _titulo(ws, 'A11', 'DATOS DEL PERIODO — rellena estas casillas')
    for fila, etiqueta, valor, fmt, ayuda in ENTRADAS_KPI:
        ws.cell(row=fila, column=1, value=etiqueta)
        cel = ws.cell(row=fila, column=2, value=valor)
        cel.number_format = fmt
        aux = ws.cell(row=fila, column=4, value=ayuda)
        aux.alignment = Alignment(wrap_text=True, vertical='top')
    motor.marcar_verde(ws, 'B12:B17')
    _nota(ws, 'A19',
          'Cifras de ejemplo coherentes con el mes de enero de Coste por '
          'Categoría: 10.400 € de compras, 34.000 € de ventas y 1.450 '
          'cubiertos dan un food cost del 31,0 %, dentro del objetivo. '
          'Cámbialas por las tuyas.')
    _ancho(ws, [36, 20, 16, 62, 16])
    cambios.append(
        '07:Dashboard KPIs!B4:B9 + C4:C8 + E4:E8 + B10: la hoja DESCRIBÍA EN '
        'TEXTO cómo calcular cada KPI y no tenía ni una fórmula ni una celda '
        'de entrada (DOM-10/TEC-08/COM-11/COM-23). Food cost sobre CONSUMO, '
        'coste por cubierto sin Limpieza ni Otros, objetivo relativo al '
        'ticket medio (C5) y las 6 entradas verdes del periodo. Ronda 2: '
        'columna ESTADO con semáforo real (RD-13/RT-06/RC-13 — era el único '
        'dashboard del kit sin él y su print_area ya reservaba la columna E '
        'vacía), objetivos NUMÉRICOS y editables (RT-19) y selector de mes, '
        'porque el KPI de compras leía siempre los doce (RD-27)')


# ==========================================================================
# BONUS-09 — calculadora de punto de pedido
# ==========================================================================
CAB_CALC = [
    '#',                                 # A
    'Producto',                          # B
    'Categoría',                         # C  (DV de este grupo)
    'Consumo diario (ud/día)',           # D
    'Lead time (días)',                  # E
    'Cobertura de seguridad (días)',     # F
    'Stock de seguridad (ud)',           # G  = D × F
    'Punto de pedido (ud)',              # H  = D × E + G
    'Precio/ud (€)',                     # I
    'Vida útil (días)',                  # J
    # RD-22 · la tercera capa del tope. La cantidad sugerida estaba capada por
    # vida útil pero no por el sitio donde hay que meter el género: sugería 9
    # barriles de cerveza a un local cuyo par MAX del 01 son 6, y 17 kg de
    # pollo con un par max de 20. Cursar un pedido así significa rechazarlo en
    # la puerta por falta de cámara.
    'Stock máximo (ud) — par max del 01',   # K
    'EOQ teórica (ud)',                  # L
    'Cantidad a pedir sugerida (ud)',    # M  = MIN(EOQ, vida×0,7, stock máx.)
    'Frecuencia de pedido (días)',       # N  = M / D
    'Proveedor',                         # O
]

#: (índice, cuántas) sobre el layout de la v1.1 (A:I). 1+1+2+1+1 = 6 → A:O.
#: El (11, 1) es la columna «Stock máximo» de la ronda 2 (RD-22): va ANTES de
#: la EOQ para que las tres capas del tope se lean en orden.
INSERTS_CALC = [(3, 1), (6, 1), (9, 2), (11, 1), (13, 1)]

F_SEGURIDAD = '=IF(OR($D{f}="",$F{f}=""),"",$D{f}*$F{f})'
#: La SPEC pide `IF(OR($D="",$E=""),…)`. Se añade `$F`: si la cobertura está
#: vacía, `$G` devuelve "" y `$D*$E+$G` da #¡VALOR! en Excel, no un número.
F_PUNTO = '=IF(OR($D{f}="",$E{f}="",$F{f}=""),"",$D{f}*$E{f}+$G{f})'
#: RT-03 · la EOQ dividía por «precio × % de almacenamiento» SIN IFERROR y sin
#: guarda de cero: un precio 0 (producto de coste cero, invitación, o
#: placeholder mientras se pide tarifa) devolvía #¡DIV/0! y lo propagaba a la
#: cantidad sugerida y a la frecuencia. Y como `Parámetros!D5` es una casilla
#: VERDE editable, teclear ahí un 0 % reventaba las 30 filas a la vez. La
#: cabecera de la SPEC exige «IFERROR/doble guarda en toda división».
F_EOQ = ('=IFERROR(IF(OR($D{f}="",$D{f}<=0,$I{f}="",$I{f}<=0,'
         'Parámetros!$D$4<=0,Parámetros!$D$5<=0),"",'
         'ROUND(SQRT(2*$D{f}*365*Parámetros!$D$4/'
         '($I{f}*Parámetros!$D$5)),0)),"")')
#: RT-08 · el factor de vida útil (0,7) iba HARDCODEADO en las 30 fórmulas,
#: que es el mismo pecado que el R1 denunció en TEC-02 con el 2 y el 0,5. Y no
#: era invisible: `Parámetros` le dedicaba una fila entera que PARECÍA un
#: parámetro y no lo era. Ahora sale de `Parámetros!$D$12`.
#: RD-22 · y el MIN tiene tres términos: EOQ, vida útil y stock máximo.
F_PEDIR = ('=IFERROR(IF(OR($L{f}="",$J{f}=""),"",'
           'MIN($L{f},ROUND($D{f}*$J{f}*Parámetros!$D$12,0),'
           'IF($K{f}="",$L{f},$K{f}))),"")')
F_FRECUENCIA = ('=IF(OR($D{f}="",$D{f}<=0,$M{f}=""),"",'
                'ROUND($M{f}/$D{f},0))')

#: §4 — 8 filas de ejemplo coherentes con las 50 del 01 (mismos nombres,
#: mismas unidades, mismos precios) y con los proveedores del 02.
#:
#: Ronda 2 · RD-02/RD-21 — el kit daba TRES puntos de pedido distintos para el
#: mismo producto, y en dos casos el de aquí estaba por encima del par MÁXIMO
#: del 01: mandaba reponer cuando ya tienes más de lo que la otra plantilla te
#: permite tener. Y el consumo diario no cuadraba con el gasto mensual del
#: Top 20 del 07 para los mismos productos y los mismos precios. Ahora el
#: consumo, el lead time, la cobertura, la vida útil y el stock máximo salen
#: de `motor.CONSUMO_EJEMPLO`, la ÚNICA tabla del kit, elegida de forma que
#:     punto de pedido = consumo × (lead + cobertura) = par level del 01
#: es decir: el par level del 01 y el punto de pedido de aquí son el mismo
#: número por construcción, no por casualidad.
#: (producto, categoría, precio/ud del 01, proveedor)
_CALC_BASE = [
    ('Pechuga de pollo', 'Cárnicos', 6.50, 0),
    ('Solomillo de ternera', 'Cárnicos', 32.00, 0),
    ('Salmón fresco', 'Pescados', 14.00, 1),
    ('Leche entera', 'Lácteos', 0.85, 3),
    ('Patata', 'Verduras/Frutas', 0.85, 2),
    ('Aceite de oliva virgen extra', 'Secos/Granos', 5.80, 3),
    ('Arroz redondo', 'Secos/Granos', 1.45, 3),
    ('Cerveza de grifo (barril 30 L)', 'Bebidas Alcohólicas', 55.00, 4),
]

#: (producto, categoría, consumo/día, lead time, cobertura, precio, vida útil,
#:  stock máximo, proveedor)
CALC_EJEMPLO = [
    (nombre, categoria,
     motor.CONSUMO_EJEMPLO[nombre][0], motor.CONSUMO_EJEMPLO[nombre][1],
     motor.CONSUMO_EJEMPLO[nombre][2], precio,
     motor.CONSUMO_EJEMPLO[nombre][3], motor.CONSUMO_EJEMPLO[nombre][4],
     motor.PROVEEDORES_MARCADOS[prov])
    for nombre, categoria, precio, prov in _CALC_BASE
]

#: `Parámetros` — (fila, parámetro, rango habitual, nota).
PARAMETROS = [
    (4, 'Coste de pedido (€)', '2-5 €',
     'Lo que te cuesta CURSAR un pedido: tiempo de hacerlo, recepción, '
     'cotejo del albarán y registro de la factura.'),
    (5, 'Coste de almacenamiento (% anual)', '20-30 % anual',
     'Sobre el VALOR del producto: capital inmovilizado, cámara, seguro y '
     'la merma de lo que se estropea esperando.'),
    (6, 'Lead time cárnicos', '1-2 días', 'Entregas diarias habituales.'),
    (7, 'Lead time pescado', '1 día', 'Entrega diaria obligatoria.'),
    (8, 'Lead time secos', '3-5 días', 'Pedidos semanales.'),
    (9, 'Lead time congelados', '2-3 días', 'Pedidos 2 veces por semana.'),
    (10, 'Cobertura de seguridad (días) — columna F', '1-3 días',
     'Se teclea en DÍAS. La columna G la convierte en unidades: consumo '
     'diario × días de cobertura. Antes se pedía en días aquí y se sumaba '
     'como unidades allí, y el punto de pedido salía a la mitad.'),
    (11, 'Vida útil (días) — columna J', 'Fresco 2-5 · refrigerado 7-15 · '
     'seco 180-720',
     'Días que aguanta el producto en tu cámara desde que lo recibes.'),
    (12, 'Factor de vida útil', '0,7 (70 %)',
     'La cantidad sugerida nunca pasa de este porcentaje de lo que consumes '
     'dentro de la vida útil: el resto es el margen para un día flojo. '
     'RT-08: es una casilla verde de verdad, no un número escondido dentro '
     'de las 30 fórmulas.'),
    (13, 'Stock máximo (ud) — columna K', 'El par MAX de la plantilla 01',
     'El tercer tope de la cantidad sugerida: por mucho que digan la EOQ y '
     'la vida útil, no se pide más de lo que cabe en tu cámara. Déjalo en '
     'blanco si no quieres ese tope.'),
]

INSTRUCCIONES_09 = [
    'BONUS: Calculadora Punto de Pedido',
    'AI Chef Pro — aichef.pro',
    None,
    'Calcula cuándo hay que pedir cada producto y cuánto conviene pedir.',
    None,
    'LAS DOS FÓRMULAS:',
    'Punto de pedido = consumo diario × lead time + stock de seguridad',
    'Stock de seguridad = consumo diario × días de cobertura (por eso la '
    'cobertura se teclea en DÍAS y la columna de al lado la pasa a '
    'unidades)',
    None,
    'PESTAÑAS:',
    '1. Calculadora — una línea por producto; las columnas SIN fondo verde '
    'se calculan solas (en este kit el verde es siempre lo que escribes tú)',
    '2. Parámetros — coste de pedido, coste de almacenamiento y factor de '
    'vida útil, editables; de ahí salen las EOQ y el tope de la cantidad '
    'sugerida',
    None,
    'CÓMO SE USA:',
    'Rellena las casillas verdes: consumo diario, lead time, cobertura, '
    'precio, vida útil y proveedor.',
    'Cuando el stock caiga hasta el punto de pedido, cursa el pedido. Ni '
    'antes ni después.',
    'La EOQ teórica es la fórmula de Wilson: el lote que minimiza el coste '
    'de pedir más el de tener el género parado. Sale del precio y de los dos '
    'parámetros de la otra pestaña, no de dos números escondidos dentro de '
    'la fórmula.',
    'La cantidad sugerida CAPA esa EOQ tres veces, y manda la más pequeña '
    'de las tres: la propia EOQ, el 70 % de lo que consumes dentro de la '
    'vida útil (la EOQ ignora que la comida se estropea) y tu stock máximo '
    '(de nada sirve pedir más de lo que te cabe en cámara).',
    'El ejemplo de la leche, con los números del propio fichero: 8 L/día, 12 '
    'días de vida útil y un stock máximo de 60 L. La EOQ da 287 L. El tope '
    'por vida útil la baja a 67 L, que es el 70 % de esos 12 días — ocho '
    'días y medio de consumo, no doce: el 30 % restante es el margen para un '
    'día flojo. Y el stock máximo la deja finalmente en 60 L.',
    None,
    'Los parámetros vienen con 3 € de coste de pedido, un 25 % anual de '
    'coste de almacenamiento y un 70 % de factor de vida útil. Cámbialos por '
    'los tuyos en Parámetros: son tres celdas verdes y afectan a las 30 '
    'líneas.',
    'Las 8 primeras líneas traen datos de ejemplo coherentes con la '
    'plantilla 01 del kit: el consumo diario está elegido para que el punto '
    'de pedido de aquí sea EXACTAMENTE el par level de allí, y el stock '
    'máximo es su par MAX. Bórralos y pon los tuyos.',
    None,
    motor.VERSION_LINE,
]


def _pre_09(wb, cambios):
    ws = wb['Calculadora']
    if str(ws['C3'].value or '').startswith('Categoría'):
        return
    for idx, cuantas in INSERTS_CALC:
        for _ in range(cuantas):
            motor.insertar_columna(ws, idx)
    cambios.append('BONUS-09:Calculadora!A3:N3: +5 columnas (categoría, '
                   'cobertura de seguridad en días, precio/ud, vida útil y '
                   'cantidad a pedir sugerida) — sin precio no hay EOQ '
                   'posible y sin vida útil no hay forma de caparla '
                   '(DOM-06/TEC-02/COM-01)')


def _post_09(wb, cambios):
    _calculadora_09(wb, cambios)
    _parametros_09(wb, cambios)
    _instrucciones(wb['Instrucciones'], INSTRUCCIONES_09)
    cambios.append('BONUS-09:Instrucciones!A1:A24: reescritas — se RETIRA la '
                   'promesa del «simulador», que nunca existió como pestaña '
                   '(§1.6/§7-bis.4), y se explica por qué la cantidad '
                   'sugerida capa la EOQ')


def _calculadora_09(wb, cambios):
    ws = wb['Calculadora']
    r0, r1 = 4, 33
    _cabecera(ws, 3, CAB_CALC)
    _clonar_estilo(ws, [3, 6, 9, 10, 11, 13], r0, r1)
    _nota(ws, 'A2', NOTA_EJEMPLO + ' El consumo diario y la vida útil son los '
                                   'de TU casa: los de ejemplo son de un '
                                   'restaurante de menú de 60 comensales y '
                                   'salen de la MISMA tabla que los par '
                                   'levels de la plantilla 01, así que el '
                                   'punto de pedido de aquí y el par level de '
                                   'allí son el mismo número.')
    for i, dato in enumerate(CALC_EJEMPLO):
        f = r0 + i
        (producto, categoria, consumo, lead, cobertura, precio, vida,
         stock_max, prov) = dato
        ws.cell(row=f, column=2, value=producto)
        ws.cell(row=f, column=3, value=categoria)
        ws.cell(row=f, column=4, value=consumo)
        ws.cell(row=f, column=5, value=lead)
        ws.cell(row=f, column=6, value=cobertura)
        ws.cell(row=f, column=9, value=precio)
        ws.cell(row=f, column=10, value=vida)
        ws.cell(row=f, column=11, value=stock_max)
        ws.cell(row=f, column=15, value=prov)
    for f in range(r0, r1 + 1):
        _f(ws, 'G{}'.format(f), F_SEGURIDAD.format(f=f))
        _f(ws, 'H{}'.format(f), F_PUNTO.format(f=f))
        _f(ws, 'L{}'.format(f), F_EOQ.format(f=f))
        _f(ws, 'M{}'.format(f), F_PEDIR.format(f=f))
        _f(ws, 'N{}'.format(f), F_FRECUENCIA.format(f=f))
    for col in ('D', 'G', 'H', 'K', 'L', 'M'):
        _fmt(ws, '{c}{a}:{c}{b}'.format(c=col, a=r0, b=r1), motor.FMT_CANT)
    _limpiar_dv_c(ws)
    _dv(ws, 'C{}:C{}'.format(r0, r1), motor.CATEGORIAS,
        'Categoría no válida',
        'Las 10 categorías del kit. Son las mismas en las 9 plantillas.')
    _nota(ws, 'A35',
          'Columnas SIN fondo verde (no se tocan): stock de seguridad, punto '
          'de pedido, EOQ teórica, cantidad sugerida y frecuencia. Si la '
          'frecuencia sale en blanco es que el consumo diario está vacío o a '
          'cero, no que la fórmula falle. La cantidad sugerida lleva TRES '
          'topes: la EOQ, el 70 % de lo que consumes dentro de la vida útil y '
          'tu stock máximo — el más pequeño de los tres es el que manda.')
    _ancho(ws, [5, 30, 22, 17, 14, 18, 17, 17, 14, 14, 20, 15, 20, 18, 30])
    cambios.append(
        'BONUS-09:Calculadora!G4:N33: 150 fórmulas nuevas. EOQ parametrizada '
        'contra Parámetros!$D$4 y $D$5 —el 2 y el 0,5 iban HARDCODEADOS en '
        'las 30 fórmulas (DOM-06/TEC-02/COM-01)—, capada por vida útil (L), '
        'stock de seguridad en UNIDADES a partir de días de cobertura '
        '(G, DOM-12/TEC-14) y guarda de consumo cero en la frecuencia '
        '(N, TEC-15, que devolvía #DIV/0!). Ronda 2: IFERROR y guarda de cero '
        'en la EOQ (RT-03), el factor 0,7 sale de Parámetros!D12 y deja de ir '
        'hardcodeado en las 30 fórmulas (RT-08), columna «Stock máximo» y '
        'TERCER tope de la cantidad sugerida (RD-22) y los datos de ejemplo '
        'derivan de la misma tabla que los par levels del 01 (RD-02/RD-21)')


def _parametros_09(wb, cambios):
    ws = wb['Parámetros']
    ws['A1'] = 'PARÁMETROS DE REFERENCIA'
    # «Valor Típico» contiene «valor», clave de MONEDA en el motor: dejaría
    # «1-2 días» con formato de euros.
    _cabecera(ws, 3, ['Parámetro', 'Rango habitual', 'Notas',
                      'Se usa en la fórmula'])
    for fila, nombre, rango, nota in PARAMETROS:
        ws.cell(row=fila, column=1, value=nombre)
        ws.cell(row=fila, column=2, value=rango)
        cel = ws.cell(row=fila, column=3, value=nota)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
    ws['D4'] = 3
    ws['D4'].number_format = motor.FMT_EUR
    ws['D5'] = 0.25
    ws['D5'].number_format = motor.FMT_PCT1
    ws['D12'] = 0.7
    ws['D12'].number_format = motor.FMT_PCT1
    motor.marcar_verde(ws, 'D4:D5')
    motor.marcar_verde(ws, 'D12:D12')
    _nota(ws, 'A15',
          'Las TRES casillas verdes entran en las fórmulas: el coste de '
          'pedido, el coste de almacenamiento anual y el factor de vida '
          'útil. El resto de la tabla es referencia para que sepas qué '
          'escribir en cada columna de la Calculadora.')
    _ancho(ws, [38, 34, 62, 20])
    cambios.append(
        'BONUS-09:Parámetros!D4:D5: dos celdas verdes NUMÉRICAS (coste de '
        'pedido 3 € y 25 % anual de almacenamiento) — la hoja definía los '
        'dos costes como TEXTO («2-5 €», «20-30 % anual») mientras la EOQ '
        'llevaba 2 y 0,5 dentro de la fórmula (DOM-06/TEC-02); A10 describe '
        'ya la columna F en días (DOM-12/TEC-14)')


# ==========================================================================
# Entradas del pipeline
# ==========================================================================
def pre(wb, fname, cambios):
    """Sólo inserción de COLUMNAS: corre antes de que el motor fije rangos."""
    if fname.startswith('06-'):
        _pre_06(wb, cambios)
    elif fname.startswith('07-'):
        _pre_07(wb, cambios)
    elif fname.startswith('BONUS-09'):
        _pre_09(wb, cambios)


def post(wb, fname, cambios, registro=None):
    if fname.startswith('06-'):
        _post_06(wb, cambios)
    elif fname.startswith('07-'):
        _post_07(wb, cambios)
    elif fname.startswith('BONUS-09'):
        _post_09(wb, cambios)


# ==========================================================================
# Demostraciones con pycel (SPEC §5): se cambian INPUTS y se comprueba que el
# resultado se mueve en la dirección correcta. Cada bloque cita
# fichero:hoja:celda.
# ==========================================================================
EPOCA = datetime.date(1899, 12, 30)


def _serie(dias):
    """Fecha de hoy ± `dias` como NÚMERO DE SERIE de Excel. pycel restando una
    CADENA a TODAY() devuelve #VALUE!: la fecha se inyecta como serie, que es
    como la guarda el propio .xlsx."""
    return (datetime.date.today() + datetime.timedelta(days=dias)
            - EPOCA).days


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            v = xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)
    return round(v, 4) if isinstance(v, float) else v


def _set(xl, ref, valor):
    """pycel exige que la celda esté YA en el `cell_map` antes de escribirla
    (`AssertionError: Address … not found in the cell map`): se evalúa primero
    y luego se asigna."""
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            xl.evaluate(ref)
            xl.set_value(ref, valor)
            return True
        except Exception:                                        # noqa: BLE001
            return False


def _compilar(path):
    from pycel import ExcelCompiler
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        return ExcelCompiler(filename=path)


def _demo_semaforo_06(path):
    """06 — DOM-13/TEC-11/COM-09: el cuarto estado. En la v1.1 un lote
    caducado hace 10 días y otro que caduca mañana daban los dos «🔴 URGENTE»,
    y el protocolo de la hoja de al lado mandaba usar hoy todo lo rojo."""
    xl = _compilar(path)
    hoja = "'Control FIFO'"
    _ev(xl, hoja + '!L5')                       # lectura en frío
    pruebas = []
    for etiqueta, dias, tipo in (
            ('caducado hace 10 días', -10, 'Caducidad'),
            ('consumo preferente vencido hace 10 días', -10,
             'Consumo preferente'),
            ('caduca mañana', 1, 'Caducidad'),
            ('caduca en 5 días', 5, 'Caducidad'),
            ('caduca en 30 días', 30, 'Caducidad')):
        _set(xl, hoja + '!F5', tipo)
        _set(xl, hoja + '!G5', _serie(dias))
        pruebas.append({'caso': etiqueta, 'dias': dias, 'F5_tipo': tipo,
                        'K5_dias_restantes': _ev(xl, hoja + '!K5'),
                        'L5_estado': _ev(xl, hoja + '!L5')})
    estados = [p['L5_estado'] for p in pruebas]
    return {
        'ref': '06-fifo-caducidades.xlsx:Control FIFO:L5',
        'formula': F_ESTADO.format(f=5),
        'pruebas': pruebas,
        'estados_distintos': len(set(str(e) for e in estados)),
        'ok': (estados[0] != estados[1] and 'CADUCADO' in str(estados[0])
               and 'REVISAR' in str(estados[1])
               and len(set(str(e) for e in estados)) == 5),
        'nota': 'la v1.1 devolvía «🔴 URGENTE» en los tres primeros casos: '
                'no distinguía caducado de «caduca mañana» ni caducidad de '
                'consumo preferente',
    }


#: RD-07 · las seis primeras filas del Control FIFO ya llevan lotes de
#: ejemplo, así que las demostraciones trabajan sobre una fila LIBRE: escribir
#: encima de un lote sembrado dejaba los `set_value` sin efecto (pycel no
#: reevaluaba) y las tres pruebas devolvían el estado del lote de ejemplo.
FILA_DEMO_06 = 11


def _demo_apertura_06(path):
    """06 — DOM-26: la fecha límite efectiva se queda con la MÁS CERCANA. Un
    envase con caducidad a 20 días abierto hoy con 3 días de vida útil pasa a
    🟡 PRÓXIMO, no se queda en 🟢 OK."""
    xl = _compilar(path)
    hoja = "'Control FIFO'"
    f = FILA_DEMO_06
    for col in 'JKLM':
        _ev(xl, '{}!{}{}'.format(hoja, col, f))
    _set(xl, '{}!F{}'.format(hoja, f), 'Caducidad')
    _set(xl, '{}!G{}'.format(hoja, f), _serie(20))
    _set(xl, '{}!E{}'.format(hoja, f), _serie(-6))
    def _lee(caso):
        return {'caso': caso,
                'K_dias_restantes': _ev(xl, '{}!K{}'.format(hoja, f)),
                'L_estado': _ev(xl, '{}!L{}'.format(hoja, f)),
                'M_dias_en_almacen': _ev(xl, '{}!M{}'.format(hoja, f))}
    pruebas = [_lee('sin abrir (caducidad a 20 días)')]
    _set(xl, '{}!H{}'.format(hoja, f), _serie(0))
    _set(xl, '{}!I{}'.format(hoja, f), 3)
    pruebas.append(_lee('abierto hoy, 3 días de vida útil tras abrir'))
    _set(xl, '{}!I{}'.format(hoja, f), 1)
    pruebas.append(_lee('abierto hoy, 1 día de vida útil tras abrir'))
    return {
        'ref': '06-fifo-caducidades.xlsx:Control FIFO:J{0} · K{0} · L{0} · '
               'M{0}'.format(f),
        'formula': 'J ' + F_LIMITE.format(f=f) + ' · M '
                   + F_ALMACEN.format(f=f),
        'pruebas': pruebas,
        'ok': (pruebas[0]['K_dias_restantes'] == 20
               and pruebas[1]['K_dias_restantes'] == 3
               and pruebas[2]['K_dias_restantes'] == 1
               and pruebas[0]['M_dias_en_almacen'] == 6),
        'nota': 'en la v1.1 no existían ni la apertura ni la vida útil tras '
                'abrir, y «Fecha Entrada» no aparecía en ninguna de las 100 '
                'fórmulas del libro (DOM-24)',
    }


def _demo_alertas_06(path):
    """06 — DOM-25/DOM-27: el valor en riesgo y los contadores de la hoja de
    alertas se mueven con la cantidad y el precio del lote.

    Ronda 2 · RD-14 — la hoja separa ahora el VALOR REALMENTE EN RIESGO
    (fila 9, todo menos 🟢 OK) del valor total del stock registrado (fila 10).
    Antes la fila del pie presentaba el valor de TODO el stock bajo la
    cabecera «Valor en riesgo (€)» y desmentía a las cinco de arriba.
    """
    xl = _compilar(path)
    fifo, al = "'Control FIFO'", "'Alertas Caducidad'"
    f = FILA_DEMO_06
    base = {'B4_caducados': _ev(xl, al + '!B4'),
            'C4_valor_caducado': _ev(xl, al + '!C4'),
            'B9_lotes_en_riesgo': _ev(xl, al + '!B9'),
            'C9_valor_realmente_en_riesgo': _ev(xl, al + '!C9'),
            'B10_lotes_totales': _ev(xl, al + '!B10'),
            'C10_valor_total_del_stock': _ev(xl, al + '!C10'),
            'C8_valor_de_los_OK': _ev(xl, al + '!C8')}
    for col in 'LQ':
        _ev(xl, '{}!{}{}'.format(fifo, col, f))
    _set(xl, '{}!B{}'.format(fifo, f), 'Solomillo de ternera')
    _set(xl, '{}!F{}'.format(fifo, f), 'Caducidad')
    _set(xl, '{}!G{}'.format(fifo, f), _serie(-2))
    _set(xl, '{}!N{}'.format(fifo, f), 8)
    _set(xl, '{}!P{}'.format(fifo, f), 32)
    con_lote = {'L_estado': _ev(xl, '{}!L{}'.format(fifo, f)),
                'Q_valor_en_riesgo': _ev(xl, '{}!Q{}'.format(fifo, f)),
                'B4_caducados': _ev(xl, al + '!B4'),
                'C4_valor_caducado': _ev(xl, al + '!C4'),
                'B9_lotes_en_riesgo': _ev(xl, al + '!B9'),
                'C9_valor_realmente_en_riesgo': _ev(xl, al + '!C9'),
                'B10_lotes_totales': _ev(xl, al + '!B10'),
                'C10_valor_total_del_stock': _ev(xl, al + '!C10')}
    return {
        'ref': '06-fifo-caducidades.xlsx:Control FIFO:Q{} + '
               'Alertas Caducidad:B4 · C4 · B9 · C9 · B10 · C10'.format(f),
        'formula': 'Q ' + F_RIESGO.format(f=f)
                   + ' · C4 =SUMIF(\'Control FIFO\'!$L$5:$L$54,'
                     '"*CADUCADO*",\'Control FIFO\'!$Q$5:$Q$54) · '
                     'C9 =SUM($C$4:$C$7) · C10 =SUM(\'Control FIFO\'!$Q$5:$Q$54)',
        'con_los_seis_lotes_de_ejemplo': base,
        'anadiendo_un_lote_caducado_de_8_kg_a_32': con_lote,
        'ok': (con_lote['Q_valor_en_riesgo'] == 256
               and con_lote['B4_caducados'] == base['B4_caducados'] + 1
               and con_lote['C4_valor_caducado']
               == base['C4_valor_caducado'] + 256
               and con_lote['B10_lotes_totales']
               == base['B10_lotes_totales'] + 1
               and base['C9_valor_realmente_en_riesgo']
               < base['C10_valor_total_del_stock']),
        'nota': 'la hoja de alertas tenía 0 fórmulas y el Control FIFO no '
                'registraba cantidad: 200 g de perejil y 8 kg de solomillo '
                'generaban exactamente la misma alerta (DOM-25). Y la fila '
                'del pie llamaba «valor en riesgo» al valor de TODO el stock '
                '(RD-14): ahora el valor realmente en riesgo es ESTRICTAMENTE '
                'menor que el total, porque hay un lote 🟢 OK',
    }


def _demo_kpis_07(path):
    """07 — DOM-10/COM-11/COM-23: el food cost se calcula sobre CONSUMO y se
    mueve con las ventas; el coste por cubierto excluye Limpieza y Otros."""
    xl = _compilar(path)
    d = "'Dashboard KPIs'"
    lectura = {'B4_food_cost': _ev(xl, d + '!B4'),
               'B5_coste_cubierto': _ev(xl, d + '!B5'),
               'B6_compras': _ev(xl, d + '!B6'),
               'B7_variacion': _ev(xl, d + '!B7'),
               'C5_objetivo': _ev(xl, d + '!C5')}
    pruebas = [dict(caso='datos de ejemplo (34.000 € de ventas, 1.450 '
                         'cubiertos)', **lectura)]
    for etiqueta, ventas in (('ventas a la mitad (17.000 €)', 17000),
                             ('ventas al doble (68.000 €)', 68000)):
        _set(xl, d + '!B12', ventas)
        pruebas.append({'caso': etiqueta, 'B12_ventas': ventas,
                        'B4_food_cost': _ev(xl, d + '!B4')})
    _set(xl, d + '!B12', 34000)
    _set(xl, d + '!B13', 2900)
    pruebas.append({'caso': 'el doble de cubiertos (2.900)',
                    'B5_coste_cubierto': _ev(xl, d + '!B5')})
    _set(xl, d + '!B17', 40)
    pruebas.append({'caso': 'ticket medio 40 € (objetivo = 30 % del ticket)',
                    'C5_objetivo': _ev(xl, d + '!C5')})
    return {
        'ref': '07-analisis-costes-compras.xlsx:Dashboard KPIs:B4 · B5 · '
               'B6 · B7 · C5',
        'formulas': {
            'B4': '=IFERROR(($B$14+$B$6-$B$15)/$B$12,"")',
            'B5': "=IFERROR(($B$14+$B$6-$B$15-'Coste por Categoría'!$N$12"
                  "-'Coste por Categoría'!$N$13)/$B$13,\"\")",
            'B6': "='Coste por Categoría'!$N$14",
            'C5': '=IFERROR($B$17*0.3,"")'},
        'pruebas': pruebas,
        'ok': (lectura['B6_compras'] == 10400
               and abs(lectura['B4_food_cost'] - 0.3103) < 0.001
               and pruebas[1]['B4_food_cost'] > lectura['B4_food_cost']
               and pruebas[2]['B4_food_cost'] < lectura['B4_food_cost']
               and pruebas[4]['C5_objetivo'] == 12.0),
        'nota': 'las dos hojas que dan nombre al entregable en el dashboard '
                'post-pago no tenían UNA SOLA FÓRMULA, y la columna D '
                'describía en texto cómo hacer la cuenta a mano (COM-11); '
                'el objetivo de coste por cubierto era «< 4,50 €», que sólo '
                'vale para un ticket de 15 € (DOM-31/COM-23)',
    }


def _demo_variacion_07(path):
    """07 — DOM-29: la variación de precio existe (la v1.1 no tenía columna de
    precio anterior) y arrastra el contador y el nombre del dashboard."""
    xl = _compilar(path)
    t, d = "'Top 20 Productos'", "'Dashboard KPIs'"
    base = {'J4_solomillo': _ev(xl, t + '!J4'),
            'J9_aceite': _ev(xl, t + '!J9'),
            'G4_pct_top20': _ev(xl, t + '!G4'),
            'E24_total': _ev(xl, t + '!E24'),
            'L_top1': _ev(xl, '{}!L{}'.format(t, FILA_TOP5 + 1)),
            'M_top1_gasto': _ev(xl, '{}!M{}'.format(t, FILA_TOP5 + 1)),
            'B8_con_subida': _ev(xl, d + '!B8'),
            'B9_mayor_subida': _ev(xl, d + '!B9'),
            'E4_estado_del_food_cost': _ev(xl, d + '!E4')}
    _set(xl, t + '!I4', 40.00)                 # solomillo 30,50 → 40,00
    subido = {'J4_solomillo': _ev(xl, t + '!J4'),
              'B8_con_subida': _ev(xl, d + '!B8'),
              'B9_mayor_subida': _ev(xl, d + '!B9')}
    _set(xl, t + '!E4', 5000)                  # y pasa a ser el nº 1 por gasto
    reordenado = {'E24_total': _ev(xl, t + '!E24'),
                  'G4_pct_top20': _ev(xl, t + '!G4'),
                  'L_top1': _ev(xl, '{}!L{}'.format(t, FILA_TOP5 + 1)),
                  'M_top1_gasto': _ev(xl, '{}!M{}'.format(t, FILA_TOP5 + 1))}
    return {
        'ref': '07-analisis-costes-compras.xlsx:Top 20 Productos:J4 · G4 · '
               'L{0}:M{1} + Dashboard KPIs:B8 · B9 · E4'.format(
                   FILA_TOP5 + 1, FILA_TOP5 + 5),
        'formulas': {
            'J4': '=IF(OR($H4="",$I4=""),"",$I4/$H4-1)',
            'G4': '=IFERROR($E4/$E$24,"")',
            'L{}'.format(FILA_TOP5 + 1):
                '=IFERROR(INDEX($B$4:$B$23,MATCH(LARGE($E$4:$E$23,1),'
                '$E$4:$E$23,0)),"")',
            'B8': '=COUNTIF(\'Top 20 Productos\'!$J$4:$J$23,">0.05")'},
        'con_los_ejemplos': base,
        'subiendo_el_solomillo_a_40': subido,
        'y_pasando_su_gasto_a_5000': reordenado,
        'ok': (base['B8_con_subida'] == 2
               and base['B9_mayor_subida'] == 'Aceite de oliva virgen extra'
               and base['L_top1'] == 'Solomillo de ternera'
               and subido['B8_con_subida'] == 3
               and subido['B9_mayor_subida'] == 'Solomillo de ternera'
               and reordenado['L_top1'] == 'Solomillo de ternera'
               and reordenado['M_top1_gasto'] == 5000
               and base['E4_estado_del_food_cost'] not in ('', None)),
        'nota': 'la landing promete «alertas de variación de precios (>5%)» '
                'y en la v1.1 la columna «Variación vs Anterior» estaba '
                'vacía en las 20 filas y era imposible de calcular: no '
                'existía «Precio Anterior» (DOM-29/TEC-08). Ronda 2: el '
                'bloque Top 5 se ha movido de L4:M8 a L{}:M{} porque vivía '
                'DENTRO del rango de datos que la propia hoja mandaba '
                'ordenar (RD-08/RT-11), y el Dashboard estrena columna de '
                'Estado (RD-13/RT-06/RC-13)'.format(FILA_TOP5 + 1,
                                                    FILA_TOP5 + 5),
    }


def _demo_eoq_09(path):
    """BONUS-09 — DOM-06/TEC-02/COM-01: la EOQ ya sale del precio y de los dos
    parámetros editables, y la cantidad sugerida la capa por vida útil.

    Ronda 2 · se añaden los tres casos que refutaron la v2.0: el precio cero
    que devolvía #¡DIV/0! y lo propagaba a tres columnas (RT-03), el factor de
    vida útil que ahora vive en `Parámetros!D12` en vez de ir hardcodeado 30
    veces (RT-08) y el TERCER tope, el stock máximo del 01 (RD-22).
    """
    xl = _compilar(path)
    c, p = "'Calculadora'", "'Parámetros'"
    fila_leche = 7                              # Leche entera, 8 L/día
    lectura = {
        'L7_eoq_leche': _ev(xl, c + '!L{}'.format(fila_leche)),
        'M7_cantidad_sugerida': _ev(xl, c + '!M{}'.format(fila_leche)),
        'N7_frecuencia_dias': _ev(xl, c + '!N{}'.format(fila_leche)),
        'K7_stock_maximo': _ev(xl, c + '!K{}'.format(fila_leche)),
        'L6_eoq_salmon': _ev(xl, c + '!L6'),
        'M6_cantidad_salmon': _ev(xl, c + '!M6'),
        'N6_frecuencia_salmon': _ev(xl, c + '!N6'),
    }
    pruebas = [dict(caso='datos de ejemplo', **lectura)]
    _set(xl, c + '!I6', 28.00)                  # salmón al doble de precio
    pruebas.append({'caso': 'salmón al doble de precio (28 €/kg): la EOQ baja',
                    'L6_eoq_salmon': _ev(xl, c + '!L6'),
                    'M6_cantidad_salmon': _ev(xl, c + '!M6')})
    _set(xl, c + '!I6', 14.00)
    _set(xl, p + '!D4', 30.00)                  # coste de pedido ×10
    pruebas.append({'caso': 'coste de pedido 30 € en Parámetros!D4: la EOQ '
                            'sube en TODAS las líneas',
                    'L6_eoq_salmon': _ev(xl, c + '!L6'),
                    'L7_eoq_leche': _ev(xl, c + '!L{}'.format(fila_leche))})
    _set(xl, p + '!D4', 3.00)
    # RT-08 · el factor de vida útil ya es un PARÁMETRO: cambiarlo mueve las
    # 30 líneas. En la v2.0 estaba escrito 30 veces dentro de la fórmula.
    _set(xl, p + '!D12', 0.4)
    pruebas.append({'caso': 'factor de vida útil al 40 % en Parámetros!D12 '
                            '(RT-08): la cantidad sugerida baja sin tocar '
                            'ninguna fórmula',
                    'M7_cantidad_sugerida':
                        _ev(xl, c + '!M{}'.format(fila_leche)),
                    'M6_cantidad_salmon': _ev(xl, c + '!M6')})
    _set(xl, p + '!D12', 0.7)
    _set(xl, c + '!J{}'.format(fila_leche), 2)  # leche con 2 días de vida
    pruebas.append({'caso': 'leche con 2 días de vida útil: el tope se come '
                            'la EOQ',
                    'L7_eoq_leche': _ev(xl, c + '!L{}'.format(fila_leche)),
                    'M7_cantidad_sugerida':
                        _ev(xl, c + '!M{}'.format(fila_leche)),
                    'N7_frecuencia_dias':
                        _ev(xl, c + '!N{}'.format(fila_leche))})
    _set(xl, c + '!J{}'.format(fila_leche), 12)
    # RT-03 · precio 0: la v2.0 devolvía #¡DIV/0! y lo propagaba a la cantidad
    # sugerida y a la frecuencia.
    _set(xl, c + '!I{}'.format(fila_leche), 0)
    div0 = {'caso': 'precio 0 €/ud (RT-03): sin #¡DIV/0!',
            'L7_eoq_leche': _ev(xl, c + '!L{}'.format(fila_leche)),
            'M7_cantidad_sugerida':
                _ev(xl, c + '!M{}'.format(fila_leche)),
            'N7_frecuencia_dias':
                _ev(xl, c + '!N{}'.format(fila_leche))}
    pruebas.append(div0)
    _set(xl, c + '!I{}'.format(fila_leche), 0.85)
    # RT-03 bis · y el 0 % de almacenamiento tecleado en la casilla VERDE.
    _set(xl, p + '!D5', 0)
    cero_pct = {'caso': '0 % de coste de almacenamiento en la casilla verde '
                        'Parámetros!D5 (RT-03): sin #¡DIV/0! en las 30 filas',
                'L6_eoq_salmon': _ev(xl, c + '!L6'),
                'L7_eoq_leche': _ev(xl, c + '!L{}'.format(fila_leche))}
    pruebas.append(cero_pct)
    _set(xl, p + '!D5', 0.25)
    return {
        'ref': 'BONUS-09-calculadora-punto-pedido.xlsx:Calculadora:L7 · M7 · '
               'N7 (Leche entera) y L6 · M6 (Salmón fresco)',
        'formulas': {'L': F_EOQ.format(f=4), 'M': F_PEDIR.format(f=4),
                     'N': F_FRECUENCIA.format(f=4)},
        'pruebas': pruebas,
        'ok': (lectura['L7_eoq_leche'] == 287
               and lectura['M7_cantidad_sugerida'] == 60
               and lectura['K7_stock_maximo'] == 60
               and lectura['N7_frecuencia_dias'] == 8
               and lectura['M6_cantidad_salmon'] == 4
               and pruebas[1]['L6_eoq_salmon'] < lectura['L6_eoq_salmon']
               and pruebas[2]['L6_eoq_salmon'] > lectura['L6_eoq_salmon']
               and pruebas[3]['M6_cantidad_salmon']
               < lectura['M6_cantidad_salmon']
               and str(div0['L7_eoq_leche']) != '#DIV/0!'
               and str(div0['M7_cantidad_sugerida']) != '#DIV/0!'
               and str(cero_pct['L6_eoq_salmon']) != '#DIV/0!'),
        'nota': 'la v1.1 mandaba 454 L de leche de golpe (la EOQ pura) sobre '
                'un producto que aguanta 12 días. La v2.0 lo capó por vida '
                'útil pero no por el sitio donde hay que meterlo: ahora la '
                'leche se queda en 60 L, que es el par MÁXIMO de la plantilla '
                '01 (RD-22), y la EOQ ya no revienta con un precio 0 ni con '
                'un 0 % de almacenamiento (RT-03)',
    }


def _demo_stock_seguridad_09(path):
    """BONUS-09 — DOM-12/TEC-14: el stock de seguridad se teclea en DÍAS y la
    hoja lo convierte a unidades. La v1.1 lo sumaba en unidades mientras
    «Parámetros» lo definía en días: el punto de pedido salía a la mitad."""
    xl = _compilar(path)
    c = "'Calculadora'"
    fila = 7                                    # Leche entera: 8 L/día
    _ev(xl, c + '!H{}'.format(fila))
    pruebas = []
    for cobertura in (1, 2, 3):
        _set(xl, c + '!F{}'.format(fila), cobertura)
        pruebas.append({
            'caso': '{} día(s) de cobertura'.format(cobertura),
            'F_cobertura_dias': cobertura,
            'G_stock_seguridad_ud': _ev(xl, c + '!G{}'.format(fila)),
            'H_punto_de_pedido_ud': _ev(xl, c + '!H{}'.format(fila))})
    _set(xl, c + '!F{}'.format(fila), 1)
    return {
        'ref': 'BONUS-09-calculadora-punto-pedido.xlsx:Calculadora:G7 · H7',
        'formulas': {'G': F_SEGURIDAD.format(f=4), 'H': F_PUNTO.format(f=4)},
        'consumo_diario_D7': _ev(xl, c + '!D{}'.format(fila)),
        'lead_time_E7': _ev(xl, c + '!E{}'.format(fila)),
        'pruebas': pruebas,
        'ok': (pruebas[1]['G_stock_seguridad_ud'] == 16
               and pruebas[1]['H_punto_de_pedido_ud'] == 32
               and pruebas[0]['H_punto_de_pedido_ud']
               < pruebas[1]['H_punto_de_pedido_ud']
               < pruebas[2]['H_punto_de_pedido_ud']),
        'nota': 'con 8 L/día y 2 días de lead time, la v1.1 daba a quien '
                'escribía «2» siguiendo su propia hoja Parámetros un punto '
                'de pedido a la mitad; con 1 día de cobertura —el del '
                'ejemplo— salen 24 L, que es EXACTAMENTE el par level de la '
                'leche en la plantilla 01 (RD-02: los dos ficheros dan el '
                'mismo número por construcción)',
    }


def demos(carpeta, origen=None):
    """Bloques de demostración de este grupo. `main.py` los mezcla con los
    suyos en el informe."""
    p06 = os.path.join(carpeta, FICHEROS[0])
    p07 = os.path.join(carpeta, FICHEROS[1])
    p09 = os.path.join(carpeta, FICHEROS[2])
    fuera = {}
    for clave, path, fn in (
            ('c_semaforo_cuatro_estados_06', p06, _demo_semaforo_06),
            ('c_fecha_limite_y_apertura_06', p06, _demo_apertura_06),
            ('c_valor_en_riesgo_y_alertas_06', p06, _demo_alertas_06),
            ('c_kpis_reales_07', p07, _demo_kpis_07),
            ('c_variacion_precio_top20_07', p07, _demo_variacion_07),
            ('c_eoq_parametrizada_y_capada_09', p09, _demo_eoq_09),
            ('c_stock_seguridad_en_unidades_09', p09, _demo_stock_seguridad_09),
    ):
        if not os.path.isfile(path):
            continue
        try:
            fuera[clave] = fn(path)
        except Exception as e:                                   # noqa: BLE001
            fuera[clave] = {'error': '{}: {}'.format(type(e).__name__, e)}
    return fuera
