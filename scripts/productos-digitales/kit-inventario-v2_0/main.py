#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 del Kit de Control de Inventario
y Compras.

    python3 main.py --dry-run [--json informe.json]

`--solo` construye el kit ENTERO por defecto (`motor,a,b,c`). Se puede pedir
una pasada parcial (`--solo motor`, `--solo a,b`), pero entonces el informe
la marca como PARCIAL y devuelve exit != 0: el resultado no es la v2.0 y no
se puede dar por bueno. Antes el default era `motor`, así que este mismo
comando producía una copia con 0 fórmulas nuevas y cerraba con «TODO VERDE»
(RD-29/RC-11).

REGLA DURA: `astro-site/public/dl/kit-inventario/` NO se toca. `--dry-run`
regenera una copia en el scratchpad y trabaja allí. Sin `--dry-run` el script
ABORTA salvo que se le pase `KIT_INVENTARIO_APPLY=1` en el entorno: la
ejecución real la hace el orquestador cuando la ronda 2 dé verde
(`kit-inventario-v2-SPEC.md`, cabecera).

LOS MÓDULOS DE GRUPO PUEDEN NO EXISTIR. `--solo motor` corre sólo el §1;
`--solo a,b` carga los que encuentre y avisa —sin fallar— de los que no. Así
este orquestador es útil desde el primer commit y no hay que tocarlo cuando
`grupo_a.py`, `grupo_b.py` y `grupo_c.py` aterricen.

Qué hace, en este orden:
  1. Copia de trabajo (dry-run) desde `dl/kit-inventario`.
  2. Por fichero: grupo.pre → motor.aplicar (§1) → grupo.post → motor.cerrar.
     Las columnas se insertan en `pre()` (antes de que el motor fije rangos);
     las filas se AÑADEN en `post()`.
  3. Idempotencia: repite el pipeline sobre un CLON del resultado y compara
     celda a celda (valor, formato, relleno, bloqueo, merges, DV, CF,
     protección, print_area). Debe dar 0 diferencias.
  4. `inject_cache.py` sobre los ficheros tocados. SIEMPRE al final: cualquier
     guardado posterior de openpyxl borraría el valor cacheado y el cliente
     vería las celdas de resultado en blanco en el visor del móvil.
  5. Verificación `data_only` de todas las fórmulas nuevas del registro. La
     celda sin valor cacheado no se exime por su texto: se le pregunta a pycel
     cuánto vale y sólo se acepta el blanco si pycel también da vacío (RT-17).
  6. `censo-entregables.py --only <carpeta> --fail --quiet`.
  7. Gate §1.6: toda pestaña entrecomillada en Instrucciones tiene que
     existir. BLOQUEA siempre (un gate que no puede suspender no es un gate).
  8. Casos de demostración con pycel (SPEC §5): el semáforo de stock, el
     semáforo FIFO de cuatro estados, el IVA por categoría y la doble guarda.
     pycel NO implementa `COUNTA` ni `MODE` (medido el 2026-08-23:
     `UnknownFunction`); las demostraciones usan `COUNTIF`/`INDEX`/`MATCH`.

Térmica: todo en SERIE, un python a la vez. No hay builds ni navegador.
"""
import argparse
import contextlib
import datetime
import importlib
import json
import logging
import os
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
ORIGEN = os.path.join(ROOT, 'astro-site', 'public', 'dl', 'kit-inventario')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    '7340312f-b4fe-4aa1-b254-4b0c17c8375f/scratchpad')
DESTINO = os.path.join(SCRATCH, 'dryrun-v2', 'kit-inventario')
IDEM = os.path.join(SCRATCH, 'dryrun-v2', 'kit-inventario-idem')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')

SPEC = 'scripts/productos-digitales/kit-inventario-v2-SPEC.md §1 (motor)'
ENV_APPLY = 'KIT_INVENTARIO_APPLY'


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Copia de trabajo
# ==========================================================================
def preparar_copia():
    if not os.path.isdir(ORIGEN):
        raise SystemExit('No existe el origen: ' + ORIGEN)
    if os.path.isdir(DESTINO):
        shutil.rmtree(DESTINO)
    os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
    shutil.copytree(ORIGEN, DESTINO)
    log('  copia de trabajo regenerada: ' + DESTINO)


# ==========================================================================
# Huella comparable e idempotencia
# ==========================================================================
def digest(path):
    """Huella de un .xlsx: valores, formatos, relleno, bloqueo, merges, DV, CF,
    protección y área de impresión. Es lo que compara la 2.ª pasada."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked))
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted('{}:{}:{}'.format(dv.type, dv.formula1, dv.sqref)
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted('{}:{}'.format(cf.sqref, len(cf.rules))
                         for cf in ws.conditional_formatting),
            'prot': bool(ws.protection.sheet),
            'area': str(ws.print_area),
        }
    fuera['__props__'] = {
        'celdas': {'title': (repr(wb.properties.title), '', None, False),
                   'subject': (repr(wb.properties.subject), '', None, False),
                   'keywords': (repr(wb.properties.keywords), '', None, False),
                   'creator': (repr(wb.properties.creator), '', None, False)},
        'merges': [], 'dv': [], 'cf': [], 'prot': False, 'area': 'None'}
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append('{}:{}: hoja sólo en una pasada'.format(fichero, hoja))
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'prot', 'area'):
            if ha[k] != hb[k]:
                fuera.append('{}:{}: cambia {} ({} → {})'
                             .format(fichero, hoja, k, ha[k], hb[k]))
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append('{}:{}!{}: {} → {}'
                             .format(fichero, hoja, coord, ca.get(coord),
                                     cb.get(coord)))
    return fuera


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, grupos, informe_global):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    cambios, registro_grupo = [], []
    motor.REGISTRO = []

    # Un grupo puede declarar `PROPIOS`: ficheros donde el motor §1 no aplica
    # y el grupo hace el trabajo entero (llamando él mismo a `motor.cerrar`).
    propio = any(fname in getattr(g, 'PROPIOS', []) for g in grupos)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []) and hasattr(g, 'pre'):
            g.pre(wb, fname, cambios)

    if not propio:
        motor.aplicar(wb, fname, cambios)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []) and hasattr(g, 'post'):
            g.post(wb, fname, cambios, registro_grupo)

    gates = {}
    if not propio:
        gates = motor.cerrar(wb, fname, cambios)
    wb.save(path)

    registro = list(motor.REGISTRO)
    registro += [(h, c, f) for h, c, f in registro_grupo
                 if isinstance(f, str) and f.startswith('=')]
    informe_global.append({'fichero': fname, 'cambios': cambios,
                           'formulas_nuevas': len(registro), 'gates': gates})
    return registro, gates


def ficheros_a_tocar(grupos):
    nombres = list(motor.FICHEROS)
    for g in grupos:
        for f in getattr(g, 'FICHEROS', []):
            if f not in nombres:
                nombres.append(f)
    return nombres


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    """SIEMPRE al final del pipeline: `inject_cache.py` escribe el valor
    cacheado directamente en el XML del zip, y cualquier `wb.save()` posterior
    de openpyxl lo borraría sin avisar."""
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        fuera.append((n, r.stdout.strip(), r.returncode))
        log('    ' + (r.stdout.strip() or r.stderr.strip()[-200:]))
        if r.returncode != 0:
            log('    ERROR inject_cache: ' + r.stderr[-400:])
    return fuera


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA debe tener valor cacheado.

    RT-17 · la versión anterior EXIMÍA de comprobación a toda fórmula cuyo
    TEXTO contuviera una cadena vacía, que es el patrón de guarda de
    prácticamente todo el kit: en la pasada completa eran 1.425 de 1.725
    (82,6 %), así que el «fallos: 0» que sostenía el verde medía el 17 % del
    registro y ninguna de las fórmulas que la ronda 2 refutó habría sido
    detectada por él.

    Ahora la celda sin valor cacheado NO se exime por su texto: se le pregunta
    a pycel cuánto vale de verdad. Si pycel dice que vale la cadena vacía, el
    blanco es correcto y se cuenta como `vacias_verificadas`; si pycel devuelve
    cualquier otra cosa, es una fórmula que SÍ tenía valor y se quedó sin
    cachear — el cliente la vería en blanco en el visor del móvil — y va a
    `fallos`. Sólo se exime lo que pycel no sabe evaluar (`COUNTA`, `MODE`…),
    y eso se dice por su nombre.
    """
    fallos, vacias, sinpycel, ok, no_evaluables = [], 0, 0, 0, 0
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wbv = openpyxl.load_workbook(path, data_only=True)
        pendientes = []
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append('{}:{}!{}: hoja ausente'.format(fname, hoja,
                                                              coord))
                continue
            v = wbv[hoja][coord].value
            if v is not None:
                ok += 1
            elif any(f in formula.upper() for f in ('COUNTA(', 'MODE(',
                                                    'PMT(', 'IRR(')):
                sinpycel += 1
            else:
                pendientes.append((hoja, coord, formula))
        if not pendientes:
            continue
        try:
            xl = _pycel(path)
        except Exception as e:                                   # noqa: BLE001
            fallos.append('{}: no se pudo compilar con pycel para verificar '
                          'las {} fórmulas sin cache ({})'
                          .format(fname, len(pendientes), type(e).__name__))
            continue
        for hoja, coord, formula in pendientes:
            valor = _ev(xl, "'{}'!{}".format(hoja, coord))
            if isinstance(valor, str) and valor.startswith('ERR:'):
                no_evaluables += 1
            elif valor in ('', None):
                vacias += 1
            else:
                fallos.append('{}:{}!{}: SIN CACHE pero pycel la evalúa a {!r} '
                              '({})'.format(fname, hoja, coord, valor,
                                            formula[:60]))
    return {'con_valor': ok, 'vacias_verificadas_con_pycel': vacias,
            'sin_pycel': sinpycel, 'no_evaluables_por_pycel': no_evaluables,
            'fallos': fallos}


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-1500:])
    salida = r.stdout.strip().splitlines()
    return {'exit': r.returncode, 'salida': salida[-6:],
            'stderr': r.stderr.strip().splitlines()[-12:]}


def gate_pestanas(carpeta, nombres):
    """§1.6 — toda pestaña entrecomillada en Instrucciones tiene que existir.

    NO es un fallo del motor: es el inventario de las promesas que los grupos
    tienen que cumplir (crear `03!Proveedores` y `07!Evolución Mensual`) o
    retirar (`03!Imprimible`, `BONUS-09!Simulador`). Se informa, y sólo pasa a
    `fallos` cuando el pipeline lleva los tres grupos cargados.
    """
    fuera = {}
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        malas = motor.pestanas_citadas(wb)
        if malas:
            fuera[n] = ['Instrucciones!A{}: «{}» no existe (hojas: {})'
                        .format(f, txt, ', '.join(wb.sheetnames))
                        for f, txt in malas]
    return fuera


# ==========================================================================
# Demostraciones con pycel (SPEC §5)
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)


def _set(xl, ref, valor):
    """pycel exige que la celda esté YA en el `cell_map` antes de escribirla
    (`AssertionError: Address … not found in the cell map`): se evalúa primero
    y luego se asigna."""
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            xl.evaluate(ref)
            xl.set_value(ref, valor)
            return True
        except Exception:                                        # noqa: BLE001
            return False


def demo_semaforo_stock(carpeta):
    """`01!Cocina!H`: el estado tiene que MOVERSE al bajar el stock por debajo
    del par level. Es lo que el formato condicional colorea (§1.2)."""
    path = os.path.join(carpeta, '01-inventario-stock-diario.xlsx')
    if not os.path.isfile(path):
        return None
    xl = _pycel(path)
    ref_estado = "'Cocina'!H5"
    par = _ev(xl, "'Cocina'!E5")
    pruebas = []
    if not isinstance(par, (int, float)) or not par:
        return {'ref': '01-inventario-stock-diario.xlsx:Cocina:H5',
                'nota': 'sin par level numérico en E5', 'pruebas': []}
    for etiqueta, valor in (('por debajo del par', par * 0.5),
                            ('entre par y par×1,5', par * 1.2),
                            ('holgado', par * 3)):
        _set(xl, "'Cocina'!G5", valor)
        pruebas.append({'caso': etiqueta, 'stock': valor,
                        'estado': _ev(xl, ref_estado)})
    distintos = len(set(str(p['estado']) for p in pruebas))
    return {'ref': '01-inventario-stock-diario.xlsx:Cocina:H5',
            'par_level': par, 'pruebas': pruebas,
            'estados_distintos': distintos,
            'ok': distintos >= 2}


def demo_semaforo_fifo(carpeta):
    """`06!'Control FIFO'`: el estado tiene que cambiar con la fecha. En v1.1
    son tres estados y no distingue caducidad de consumo preferente (§4)."""
    path = os.path.join(carpeta, '06-fifo-caducidades.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Control FIFO']
    col_estado = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        if isinstance(v, str) and 'estado' in v.lower():
            col_estado = openpyxl.utils.get_column_letter(c)
            break
    if not col_estado:
        return None
    col_fecha = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        if isinstance(v, str) and 'caducidad' in v.lower():
            col_fecha = openpyxl.utils.get_column_letter(c)
            break
    xl = _pycel(path)
    hoy = datetime.date.today()
    # pycel restando una CADENA a TODAY() devuelve #VALUE!: la fecha se
    # inyecta como NÚMERO DE SERIE de Excel (días desde el 30/12/1899), que es
    # como la guarda el propio .xlsx.
    epoca = datetime.date(1899, 12, 30)
    pruebas = []
    for etiqueta, dias in (('caducado ayer', -1), ('caduca mañana', 1),
                           ('caduca en 5 días', 5), ('caduca en 30 días', 30)):
        if col_fecha:
            serie = (hoy + datetime.timedelta(days=dias) - epoca).days
            _set(xl, "'Control FIFO'!{}5".format(col_fecha), serie)
        pruebas.append({'caso': etiqueta, 'dias': dias,
                        'estado': _ev(xl, "'Control FIFO'!{}5"
                                      .format(col_estado))})
    distintos = len(set(str(p['estado']) for p in pruebas))
    return {'ref': '06-fifo-caducidades.xlsx:Control FIFO:{}5'
                   .format(col_estado),
            'columna_fecha': col_fecha, 'pruebas': pruebas,
            'estados_distintos': distintos,
            'nota': ('en v1.1 son 3 estados y el consumo preferente vencido se '
                     'marca igual que un caducado; el 4.º estado lo pone '
                     'grupo_c (§4, DOM-13)')}


def demo_iva_por_categoria(carpeta):
    """§1.3 — la tabla `03!Listas` existe y da un tipo distinto por categoría.
    Se lee del fichero (no de la constante) para demostrar que se escribió."""
    path = os.path.join(carpeta, '03-pedidos-compra.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    if motor.HOJA_LISTAS not in wb.sheetnames:
        return {'ok': False, 'nota': 'la hoja Listas no existe'}
    ws = wb[motor.HOJA_LISTAS]
    filas = []
    for r in range(2, 12):
        filas.append((ws.cell(row=r, column=1).value,
                      ws.cell(row=r, column=2).value))
    tipos = sorted(set(t for _, t in filas if t is not None))
    return {'ref': '03-pedidos-compra.xlsx:{}:A2:B11'.format(motor.HOJA_LISTAS),
            'tabla': [{'categoria': c, 'iva': t} for c, t in filas],
            'tipos_distintos': tipos,
            'ok': tipos == [4, 10, 21],
            'nota': ('hoy `Pedido Actual!G9:G28` lleva el 10 escrito a mano en '
                     'las 20 líneas: un pedido de vino sale con 11 puntos de '
                     'IVA de menos (DOM-07)')}


def demo_doble_guarda(carpeta):
    """§1.8 — la guarda de la v1.1 vigila la cantidad e ignora el precio: una
    línea sin precio vale 0,00 € y se suma como si fuera gratis (TEC-06)."""
    path = os.path.join(carpeta, '03-pedidos-compra.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Pedido Actual']
    col_sub = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=8, column=c).value
        if isinstance(v, str) and 'subtotal' in v.lower():
            col_sub = openpyxl.utils.get_column_letter(c)
            break
    if not col_sub:
        return None
    formula = ws['{}9'.format(col_sub)].value
    xl = _pycel(path)
    col_cant = openpyxl.utils.get_column_letter(
        openpyxl.utils.column_index_from_string(col_sub) - 2)
    _set(xl, "'Pedido Actual'!{}9".format(col_cant), 10)
    con_cantidad_sin_precio = _ev(xl, "'Pedido Actual'!{}9".format(col_sub))
    return {'ref': '03-pedidos-compra.xlsx:Pedido Actual:{}9'.format(col_sub),
            'formula': formula,
            'cantidad_10_sin_precio': con_cantidad_sin_precio,
            'doble_guarda': isinstance(formula, str) and 'falta coste' in formula,
            'nota': ('mientras el resultado sea 0 y no un aviso, el total del '
                     'pedido miente: la línea se suma como si fuera gratis '
                     '(§1.8, TEC-06)')}


def demo_proteccion(carpeta, nombres):
    """§1.5 — hoja a hoja: protegida sin contraseña y con las verdes abiertas."""
    fuera = []
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            verdes = sum(1 for row in ws.iter_rows() for c in row
                         if motor.es_verde(c))
            abiertas = sum(1 for row in ws.iter_rows() for c in row
                           if motor.es_verde(c) and not c.protection.locked)
            fuera.append({'ref': '{}:{}'.format(n, ws.title),
                          'protegida': bool(ws.protection.sheet),
                          'con_password': bool(ws.protection.password),
                          'verdes': verdes, 'verdes_desbloqueadas': abiertas})
    fallos = ['{}: {} verdes y sólo {} desbloqueadas'
              .format(d['ref'], d['verdes'], d['verdes_desbloqueadas'])
              for d in fuera if d['verdes'] != d['verdes_desbloqueadas']]
    fallos += ['{}: protegida CON contraseña'.format(d['ref'])
               for d in fuera if d['con_password']]
    return {'hojas': fuera, 'fallos': fallos}


# ==========================================================================
# main
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 del Kit de Control de Inventario')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    # RD-29/RC-11 · el default era 'motor', así que el comando que la propia
    # cabecera de este script documenta —y el que se pasó a los auditores—
    # producía una copia del kit con CERO fórmulas nuevas y aun así cerraba
    # con «TODO VERDE» y exit 0. Quien corriera el comando documentado, leyera
    # la última línea y diera el visto bueno, estaría aprobando un kit sin
    # construir. Ahora el default construye el kit ENTERO; para correr sólo el
    # motor hay que pedirlo a mano.
    ap.add_argument('--solo', default='motor,a,b,c',
                    help='qué aplicar: motor,a,b,c (por defecto) | motor | '
                         'a,b … (el motor va siempre)')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--sin-idempotencia', action='store_true')
    ap.add_argument('--sin-demos', action='store_true')
    args = ap.parse_args()

    if not args.dry_run and os.environ.get(ENV_APPLY) != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en ' + ORIGEN +
            ', que la SPEC declara intocable hasta que la ronda 2 dé verde. '
            'Usa --dry-run (o ' + ENV_APPLY + '=1 si eres el orquestador y ya '
            'tienes el visto bueno).')

    carpeta = DESTINO if args.dry_run else ORIGEN
    respaldo = None
    if args.dry_run:
        preparar_copia()
    else:
        # En producción se escribe IN PLACE sobre los entregables. Si algo
        # revienta a mitad del bucle, los xlsx quedan a medio transformar. El
        # respaldo va al SCRATCHPAD, nunca junto a los entregables: una carpeta
        # `.bak` dentro de `public/dl/` se publicaría.
        respaldo = os.path.join(
            SCRATCH, 'kit-inventario.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(ORIGEN, respaldo)
        log('  respaldo previo de los entregables: ' + respaldo)

    grupos, pedidos, ausentes = [], [], []
    for letra in [g.strip().lower() for g in args.solo.split(',') if g.strip()]:
        if letra == 'motor':
            continue
        # `--solo motor,grupo_a` y `--solo motor,a` son la misma cosa: sin
        # esto, la primera forma buscaría un módulo `grupo_grupo_a` y el
        # grupo se declararía «ausente» sin que nada fallara — un verde
        # falso, que es el peor resultado posible de un gate.
        if letra.startswith('grupo_'):
            letra = letra[6:]
        pedidos.append(letra)
        try:
            grupos.append(importlib.import_module('grupo_' + letra))
            log('  grupo_{} cargado'.format(letra))
        except ImportError:
            ausentes.append('grupo_' + letra)
            log('  grupo_{} no existe todavía — se omite'.format(letra))

    nombres = ficheros_a_tocar(grupos)
    informe_ficheros, registros, gates_fichero = [], {}, {}
    log('\n== 1/7 · post-proceso de {} ficheros =='.format(len(nombres)))
    for fname in nombres:
        registros[fname], gates_fichero[fname] = procesar(
            carpeta, fname, grupos, informe_ficheros)
        log('  {}: {} fórmulas nuevas'.format(fname, len(registros[fname])))

    # ---- idempotencia ---------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 2/7 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(IDEM):
            shutil.rmtree(IDEM)
        shutil.copytree(carpeta, IDEM)
        antes = dict((n, digest(os.path.join(carpeta, n))) for n in nombres)
        for fname in nombres:
            procesar(IDEM, fname, grupos, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(IDEM, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:40]}
        log('  diferencias 1.ª vs 2.ª pasada: {}'.format(len(difs)))
        for d in difs[:10]:
            log('    ' + d)

    # ---- cache ----------------------------------------------------------
    log('\n== 3/7 · inject_cache (al final del todo) ==')
    cache = inject_cache(carpeta, nombres)

    log('\n== 4/7 · verificación data_only de las fórmulas nuevas ==')
    ver = verificar_cache(carpeta, registros)
    log('  con valor: {} · "" VERIFICADA con pycel: {} · sin pycel: {} · '
        'no evaluables: {} · fallos: {}'
        .format(ver['con_valor'], ver['vacias_verificadas_con_pycel'],
                ver['sin_pycel'], ver['no_evaluables_por_pycel'],
                len(ver['fallos'])))
    for f in ver['fallos'][:10]:
        log('    ' + f)

    log('\n== 5/7 · censo-entregables --fail ==')
    cen = censo(carpeta)

    log('\n== 6/7 · gate §1.6 · pestañas citadas en Instrucciones ==')
    pest = gate_pestanas(carpeta, nombres)
    for n, lineas in pest.items():
        for l in lineas:
            log('    {}: {}'.format(n, l))
    if not pest:
        log('    0 pestañas citadas inexistentes')

    log('\n== 7/7 · demostraciones (pycel) ==')
    demos = {}
    if not args.sin_demos:
        demos = {
            'semaforo_stock_01': demo_semaforo_stock(carpeta),
            'semaforo_fifo_06': demo_semaforo_fifo(carpeta),
            'iva_por_categoria_03': demo_iva_por_categoria(carpeta),
            'doble_guarda_03': demo_doble_guarda(carpeta),
            'proteccion': demo_proteccion(carpeta, nombres),
        }
        for g in grupos:
            if hasattr(g, 'demos'):
                propias = g.demos(carpeta, ORIGEN)
                demos.update(propias)
                log('  demostraciones de {}: {} bloques'
                    .format(g.__name__, len(propias)))

    # ---- veredicto ------------------------------------------------------
    fallos = []
    if idem.get('diferencias'):
        fallos.append('idempotencia: {} diferencias'
                      .format(idem['diferencias']))
    if ver['fallos']:
        fallos.append('cache: {} fórmulas sin valor'.format(len(ver['fallos'])))
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió {}'
                      .format(cen['exit']))
    if demos.get('proteccion'):
        fallos += demos['proteccion']['fallos']
    # RD-29/RC-11 · «un gate que no puede suspender no es un gate». Antes
    # sólo bloqueaba con los tres grupos cargados, así que la pasada por
    # defecto imprimía NUEVE promesas de pestaña rotas como texto informativo
    # y terminaba en verde. Ahora bloquea SIEMPRE que un fichero cite una
    # pestaña que no existe, corran los grupos que corran: si la pasada es
    # parcial, el mensaje lo dice, pero el exit sigue siendo != 0.
    completo = all(('grupo_' + l) not in ausentes for l in ('a', 'b', 'c')) \
        and set(pedidos) >= {'a', 'b', 'c'}
    if pest:
        fallos.append(
            'gate §1.6: {} ficheros citan pestañas inexistentes{}'
            .format(len(pest), '' if completo else
                    ' (pasada PARCIAL: faltan grupos, así que el kit no está '
                    'construido — vuelve a correr con --solo motor,a,b,c)'))
    if not completo:
        fallos.append(
            'pasada PARCIAL: grupos pedidos {} · ausentes {}. El resultado NO '
            'es la v2.0 y no se puede dar por bueno.'
            .format(pedidos or ['ninguno'], ausentes or ['ninguno']))

    informe = {
        'producto': motor.PID,
        'version': motor.VERSION,
        'spec': SPEC,
        'rol': 'motor',
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'carpeta_origen': ORIGEN,
        'carpeta_trabajo': carpeta,
        'grupos_pedidos': pedidos,
        'grupos_ausentes': ausentes,
        'ficheros': informe_ficheros,
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'data_only_formulas_nuevas': ver,
            'censo_entregables': cen,
            'pestanas_citadas_inexistentes': pest,
        },
        'demostraciones': demos,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        destino_json = os.path.abspath(args.json)
        os.makedirs(os.path.dirname(destino_json), exist_ok=True)
        with open(destino_json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1)
        log('\ninforme → ' + destino_json)

    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'TODO VERDE (idempotencia, cache, censo, protección)'))
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (BÓRRALO sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
