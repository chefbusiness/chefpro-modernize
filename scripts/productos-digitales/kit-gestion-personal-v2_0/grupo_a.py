#!/usr/bin/env python3
"""
grupo_a.py — Grupo A del Kit de Gestión de Personal y Turnos v2.0:
**el tiempo de trabajo**. Cubre el §2 de `kit-gestion-personal-v2-SPEC.md`:

  · `01-cuadrante-turnos-semanal.xlsx`  — hoja nueva `Turnos`, el bloque
    auxiliar oculto `P:AP` y las **4 alertas legales** del cuadrante semanal
    (DOM-05/DOM-25/TEC-05/COM-01), más la reconstrucción del `Cuadrante
    Mensual`, que hoy es una rejilla muerta (DOM-17/TEC-23/COM-12).
  · `02-control-horas-extras.xlsx`      — cruce de medianoche y turno partido
    (DOM-03/DOM-21/TEC-01/COM-07), guarda de «H. Contratadas» vacía
    (DOM-11/TEC-09), `Resumen Mensual` que AGREGA por `SUMIF` con el recargo
    como parámetro y el contador de las 80 h/año (DOM-07/TEC-06/TEC-08/
    COM-02/COM-14) y la cita legal correcta (DOM-12/COM-19).
  · `BONUS-01-briefing-cambio-turno.xlsx` — bloques de CAJA y de TEMPERATURAS
    (DOM-29) y `wrap_text` en las observaciones (TEC-24).

Interfaz que espera `main.py` (`procesar()`): `FICHEROS`, `pre(wb, fname,
cambios)`, `post(wb, fname, cambios, registro_grupo)` y `demos(carpeta,
origen)`. NO se declara `PROPIOS`: los tres ficheros pasan por el motor §1.

REGLA DE ORO del pipeline (main.py): las COLUMNAS se insertan en `pre()`
—antes de que el motor fije rangos— y las FILAS se añaden en `post()`.

IDEMPOTENCIA. Todo lo que este módulo escribe es o bien escritura ABSOLUTA
sobre una celda, o bien una reconstrucción completa de la hoja (`Turnos` y
`Cuadrante Mensual` se borran y se vuelven a crear). Los objetos ACUMULABLES
—validaciones y formato condicional— se limpian por marca antes de
reescribirse: sin eso, la 2.ª pasada duplica cada DV y cada regla, y la
diferencia no se ve en el contenido pero sí en el fichero.

⚠ Dos cosas que el motor NO puede hacer por este grupo, y por eso van aquí:
  1. `motor.DV_LISTA['02-…']` busca «Tipo Extra» en la **columna H**. La v2.0
     mete «Pausa (h)» en E y el Tipo se va a la I, así que ese centinela ya no
     casa nunca: la DV de las 4 modalidades de hora extra (art. 35.2) la pone
     este módulo, con marca propia `grupoA` para que `motor._limpiar_dv` —que
     barre por el prefijo `kitgp-v2`— no se la lleve por delante.
  2. `motor.CF_COLUMNA['BONUS-01…']` apunta a `Briefing!D` con centinela
     «Conforme», que sólo existe en el bloque de temperaturas que crea este
     módulo y que NO está en `motor.BLOQUES`. El semáforo de ese bloque se
     pinta aquí.

Python 3.7 / openpyxl 3.1.3: sin walrus ni f-strings de depuración.
"""
import copy
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.datavalidation import DataValidation

import motor

# ==========================================================================
# Identidad del grupo
# ==========================================================================
FICHEROS = [
    '01-cuadrante-turnos-semanal.xlsx',
    '02-control-horas-extras.xlsx',
    'BONUS-01-briefing-cambio-turno.xlsx',
]

F01, F02, FB1 = FICHEROS

#: Marca de MIS validaciones. Deliberadamente NO empieza por `kitgp-v2`:
#: `motor._limpiar_dv` borra todo lo que case con ese prefijo, y una DV mía
#: barrida por el motor desaparecería sin que ningún gate lo notara (el motor
#: sólo re-escribe LAS SUYAS).
MARCA_A = 'grupoA-v2'
#: El prefijo que sí usa el motor, para poder limpiar las DV informativas que
#: `motor.parametro()` deja en las hojas que el motor luego no barre (Resumen
#: Mensual y Turnos no están en `motor.DV_LISTA`, así que nadie las limpiaría
#: y se DUPLICARÍAN en cada pasada).
MARCA_MOTOR = motor.MARCA_DV + ' · '

#: §1.3 — 30 empleados en toda hoja indexada por empleado.
CAP = motor.CAPACIDAD                       # 30
FILAS_REG = motor.FILAS_REGISTRO            # 300

#: Bloque auxiliar OCULTO del cuadrante semanal (§2). 7 + 7 + 7 + 6 = 27
#: columnas, de la P a la AP. No llevan cabecera en la fila 5 A PROPÓSITO: el
#: motor deduce formato y verde por el texto de esa fila, y una cabecera aquí
#: metería 27 columnas auxiliares en `aplicar_verde`. La etiqueta va en la
#: fila 4, que el motor no mira, para quien las muestre.
#:
#: RD-16 · el bloque arrancaba en la P (16) y se ha desplazado dos columnas
#: para que O y P queden libres: son la casilla verde «Menor de edad (S/N)» y
#: su alerta. El único aviso de minoría de edad del kit vivía en el 07 y no
#: llegaba al sitio donde se comete la infracción — en el cuadrante se podía
#: poner N (noche 23:00-07:00) o D (doble, 16 h) a un menor y ninguna de las
#: cuatro alertas decía nada, porque el 01 no conoce las fechas de nacimiento
#: y §6 prohíbe los enlaces entre libros.
AUX_HORAS = 18      # R..X  · horas de cada día
AUX_INI = 25        # Y..AE · hora de inicio
AUX_FIN = 32        # AF..AL· hora de fin
AUX_TRANS = 39      # AM..AR· las 6 transiciones entre jornadas
AUX_ULT = 44        # AR

#: RD-16 · columnas nuevas del cuadrante semanal.
COL_MENOR = 'O'     # verde · S/N, se copia del 07
COL_ALERTA_MENOR = 'P'
#: Jornada máxima de un menor: 8 h al día INCLUIDA la formación (art. 34.3 ET
#: por remisión del art. 6). No es el parámetro `jornada_diaria_max` del
#: convenio: para un menor no hay distribución irregular que lo suba.

DIAS = ('Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado',
        'Domingo')

#: Rango de la tabla de turnos, tal cual lo referencian todas las fórmulas.
JORNADA_MENOR = 8

TURNOS_TABLA = 'Turnos!$A$5:$E$12'
TURNOS_JORNADA_MAX = 'Turnos!$B$2'
TURNOS_DESCANSO = 'Turnos!$B$3'

BLANCO = Font(color='FFFFFF', bold=True, size=10)
NEGRITA = Font(bold=True)
CENTRO = Alignment(horizontal='center', vertical='center')


# ==========================================================================
# Utilidades locales
# ==========================================================================
def _f(ws, coord, formula):
    """Escribe una fórmula y la REGISTRA para que `main.py` verifique con
    `data_only` que quedó con valor cacheado (o que pycel confirma que vale la
    cadena vacía)."""
    ws[coord] = motor._reg(ws, coord, formula)
    return formula


def _cab(ws, fila, textos, col0=1):
    """Fila de cabecera con el estilo oscuro del kit."""
    for i, t in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i)
        cel.value = t
        cel.fill = PatternFill('solid', fgColor=motor.CAB)
        cel.font = BLANCO
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    ws.row_dimensions[fila].height = 30


def _titulo_bloque(ws, fila, texto, ultima_col):
    """Banda de título de un bloque, combinada de A a `ultima_col`."""
    ref = 'A{f}:{c}{f}'.format(f=fila, c=get_column_letter(ultima_col))
    if ref not in [str(m) for m in ws.merged_cells.ranges]:
        ws.merge_cells(ref)
    cel = ws.cell(row=fila, column=1)
    cel.value = texto
    cel.font = Font(bold=True, size=11)
    cel.fill = PatternFill('solid', fgColor='EEEEEE')
    cel.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[fila].height = 20


def _fila_con(ws, col, texto, desde=1):
    """Primera fila >= `desde` cuya columna `col` contiene `texto`."""
    for r in range(desde, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and texto.lower() in v.lower():
            return r
    return None


def _cola(ws, r_ult):
    """Filas de COLA de un bloque: todo lo que tiene contenido por debajo de la
    última fila de datos (totales, notas, pie). `motor.expandir_filas` las baja
    en bloque y ESTIRA las referencias a la última fila."""
    fuera = []
    for r in range(r_ult + 1, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None
               for c in range(1, ws.max_column + 1)):
            fuera.append(r)
    return tuple(fuera)


def _desverdear(ws, rango):
    """Quita el verde de un rango que pasa a ser CALCULADO, y lo bloquea.

    Hay que hacerlo AQUÍ y no dejárselo a `motor.aplicar_verde`, que sí detecta
    la columna que pasó a calculada pero la «desmarca» copiándole el relleno de
    la **columna A de su misma fila** — y en `02!Resumen Mensual` la columna A
    es el nombre del empleado, o sea VERDE. Resultado medido el 2026-08-24: las
    30 celdas del `SUMIF` de la columna B y las 30 del `SUMIFS` de la C se
    quedaban verdes y DESBLOQUEADAS, invitando al cliente a escribir encima de
    la agregación que acabamos de construir. Como `aplicar_verde` sólo entra a
    esa rama `if es_verde(cel)`, basta con haber quitado el verde antes: el
    motor las deja en paz.
    """
    n = 0
    filas = ws[rango]
    if not isinstance(filas, tuple):
        filas = ((filas,),)
    for fila in filas:
        for cel in (fila if isinstance(fila, tuple) else (fila,)):
            if motor.es_verde(cel):
                n += 1
            cel.fill = PatternFill()
            cel.protection = Protection(locked=True)
    return n


def _limpiar_mis_dv(ws, rangos=()):
    """Quita MIS validaciones (marca `grupoA-v2`), las informativas que deja
    `motor.parametro` (marca `kitgp-v2 · `) y las HEREDADAS de la v1.1 que
    pisen los rangos que voy a reescribir.

    Sin esto la 2.ª pasada acumula: `motor.parametro` añade un `DataValidation`
    nuevo en cada llamada y el motor sólo barre las hojas que él mismo gobierna
    (`Resumen Mensual` y `Turnos` no están en `motor.DV_LISTA`, así que nadie
    las limpiaría). Las heredadas hay que quitarlas porque Excel aplica la
    PRIMERA validación que encuentra: la lista vieja «Voluntaria,Obligatoria»
    seguiría ganando sobre las 4 modalidades del art. 35.2.
    """
    objetivo = set()
    for r in rangos:
        objetivo |= motor._celdas_sqref(r)
    vivas, quitadas = [], 0
    for dv in ws.data_validations.dataValidation:
        pt = getattr(dv, 'promptTitle', None) or ''
        if pt.startswith(MARCA_A) or pt.startswith(MARCA_MOTOR):
            quitadas += 1
            continue
        if objetivo and (motor._celdas_sqref(dv.sqref) & objetivo):
            quitadas += 1
            continue
        vivas.append(dv)
    ws.data_validations.dataValidation = vivas
    return quitadas


def _dv(ws, ref, valores, titulo, prompt, error, tipo='list', **extra):
    """Validación con marca propia. `showErrorMessage=True` (§ convenciones)."""
    if tipo == 'list':
        formula1 = '"{}"'.format(','.join(valores))
        if len(formula1) > 255:
            raise ValueError('DV inline de {} caracteres: {}'
                             .format(len(formula1), titulo))
        dv = DataValidation(type='list', formula1=formula1, allow_blank=True,
                            showErrorMessage=True, errorTitle=titulo,
                            error=error, errorStyle='stop',
                            showInputMessage=True,
                            promptTitle='{} · {}'.format(MARCA_A, titulo),
                            prompt=prompt)
    else:
        dv = DataValidation(type=tipo, allow_blank=True,
                            showErrorMessage=True, errorTitle=titulo,
                            error=error, errorStyle='stop',
                            showInputMessage=True,
                            promptTitle='{} · {}'.format(MARCA_A, titulo),
                            prompt=prompt, **extra)
    ws.add_data_validation(dv)
    dv.add(ref)
    return dv


def _nota_parametro(ws, fila, clave, col_rotulo=1, col_valor=2, alto=None):
    """`motor.parametro` + legibilidad: el rótulo del catálogo §1.4 es largo
    («Recargo de la hora extra (× la ordinaria) — según tu convenio») y la
    columna que lo aloja mide 16-22 caracteres, así que sin `wrap_text` se ve
    cortado y el cliente no sabe qué está editando. Se conserva el literal del
    motor —es el que explica de dónde sale el número— y se le da altura."""
    coord = motor.parametro(ws, fila, clave, col_rotulo, col_valor)
    cel = ws.cell(row=fila, column=col_rotulo)
    cel.alignment = Alignment(wrap_text=True, vertical='center')
    cel.font = NEGRITA
    if alto:
        ws.row_dimensions[fila].height = max(alto,
                                             ws.row_dimensions[fila].height
                                             or 0)
    return coord


def _instrucciones(ws, lineas):
    """Reescribe el CUERPO de `Instrucciones` entero y de forma determinista.

    `motor.linea_instrucciones` sustituye la línea que case con un patrón y, si
    no casa ninguna, la AÑADE en `max_row + 2`. Eso no es idempotente en este
    kit: el bloque de bio y versión de `motor.bio_y_version` se ancla en la
    línea de versión —que en la v1.1 es la ÚLTIMA— y pisa lo que se haya
    añadido debajo. Medido el 2026-08-24: la 1.ª pasada perdía una línea de
    cada fichero y la 2.ª la volvía a escribir dos filas más abajo → 6
    diferencias de idempotencia, una por línea nueva y otra por el `print_area`
    que crecía con ella.

    Reescribiendo el cuerpo entero no hay líneas «añadidas»: hay una lista. Las
    tres últimas son las del motor (nota de desprotección, bio y versión), que
    `bio_y_version` reconoce por su literal y reescribe EN SU SITIO.
    """
    col = motor.col_instrucciones(ws)
    estilos = {}
    for r in range(4, min(ws.max_row, 30) + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str):
            continue
        clave = 'texto' if v.startswith('▸') else 'titulo'
        if clave not in estilos:
            estilos[clave] = copy.copy(ws.cell(row=r, column=col)._style)
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=col).value = None
    fila = 4
    for texto in list(lineas) + [None, motor.NOTA_DESPROTEGER, motor.BIO,
                                 motor.VERSION_LINE]:
        if texto is None:
            fila += 1
            continue
        cel = ws.cell(row=fila, column=col)
        cel.value = texto
        est = estilos.get('texto' if texto.startswith('▸') else 'titulo')
        if est is not None:
            cel._style = copy.copy(est)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
        fila += 1
    return fila - 1


#: Las dos líneas que `motor.escribir_leyenda` reescribe en el 01 (§1.1). Se
#: toman del motor —no se copian— para que el cuerpo que escribe este módulo
#: contenga ya su literal exacto: si no estuvieran, `linea_instrucciones` las
#: añadiría DESPUÉS del bloque de bio y versión.
def _lineas_leyenda(fname):
    return [t for f, _rx, t in motor.LINEAS_LEYENDA if f == fname]


# ==========================================================================
# 01 — hoja nueva `Turnos` (§2, DOM-05/TEC-05)
# ==========================================================================
def _hoja_turnos(wb, cambios):
    """La tabla de la que salen TODAS las horas y las cuatro alertas.

    Hoy los horarios viven dentro de la fórmula monstruo de `01!Cuadrante
    Semanal!I6` —siete `IF` anidados de 700 caracteres con el 8 y el 9
    escritos dentro— y en cinco líneas de texto de `Instrucciones!B16:B20`.
    Un restaurante que abra la mañana a las 8:00 no tiene dónde decirlo.

    Las horas van como NÚMERO de 0 a 24, no como hora de Excel: las fórmulas de
    transición hacen aritmética (`24-$AD6+$X6`) y, sobre todo, pycel devuelve un
    `datetime.time` al leer una celda con formato de hora y la multiplicación
    reventaría en `inject_cache` — la celda de resultado se quedaría en blanco
    en el visor del móvil, que es justo lo que ese script existe para evitar.
    """
    idx = wb.sheetnames.index('Instrucciones') + 1 \
        if 'Instrucciones' in wb.sheetnames else 0
    if 'Turnos' in wb.sheetnames:
        idx = wb.sheetnames.index('Turnos')
        wb.remove(wb['Turnos'])
    ws = wb.create_sheet('Turnos', idx)

    for letra, ancho in (('A', 30), ('B', 34), ('C', 14), ('D', 14),
                         ('E', 10), ('F', 12)):
        ws.column_dimensions[letra].width = ancho

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Turnos y límites legales — edítalos aquí, valen para todo el kit'
    ws['A1'].font = Font(bold=True, size=13)
    ws.row_dimensions[1].height = 22

    _nota_parametro(ws, 2, 'jornada_diaria_max', 1, 2, alto=30)
    _nota_parametro(ws, 3, 'descanso_min', 1, 2, alto=30)
    ws['C2'] = 'Art. 34.3 ET'
    ws['C3'] = 'Art. 34.3 ET'
    for coord in ('C2', 'C3'):
        ws[coord].font = Font(italic=True, size=9, color='666666')

    # RT-02 · «Horas» era una CONSTANTE verde independiente de la hora de
    # inicio y de fin, y TODO el cálculo del 01 —I6, las cinco semanas del
    # mensual y la alerta N de jornada diaria— la lee por VLOOKUP con índice 5.
    # Medido en la copia dry-run: con `Turnos!D5 = 17` (mañana de 7 a 17, 10 h
    # reales, por encima del máximo de 9 h de B2), `I6` seguía en 48,0 y `N6`
    # seguía callada. Y `Instrucciones!B7` prometía justo lo contrario: «Si tu
    # mañana empieza a las 8:00, cámbialo allí y se recalcula el kit entero».
    #
    # Ahora E es CALCULADA a partir de C, D y una columna nueva de PAUSA (que
    # es lo único que impedía derivarla: el partido no es D − C). `MOD(D-C,24)`
    # resuelve el turno de noche —23 → 7 da MOD(-16,24) = 8— igual que el
    # `MOD(...,1)` del 02 resuelve el cruce de medianoche con horas de Excel.
    _cab(ws, 4, ['Código', 'Descripción', 'Hora inicio (h)', 'Hora fin (h)',
                 'Horas', 'Pausa (h)'])
    for i, cod in enumerate(motor.CODIGOS_JORNADA):
        fila = 5 + i
        ini = int(cod[2].split(':')[0])
        fin = int(cod[3].split(':')[0])
        # La pausa es lo que separa las horas EFECTIVAS del hueco entre
        # extremos: el partido va de 10 a 23 (13 h) con 4 de pausa → 9 h.
        pausa = round((fin - ini) % 24 - cod[4], 2)
        ws.cell(row=fila, column=1, value=cod[0]).alignment = CENTRO
        ws.cell(row=fila, column=1).font = NEGRITA
        ws.cell(row=fila, column=1).fill = PatternFill('solid',
                                                       fgColor=cod[-1])
        ws.cell(row=fila, column=2, value=cod[1])
        for col, valor in ((3, ini), (4, fin), (6, pausa)):
            cel = ws.cell(row=fila, column=col, value=valor)
            cel.number_format = motor.FMT_DEC2
            cel.alignment = CENTRO
        _f(ws, 'E{}'.format(fila),
           '=IF(OR($C{f}="",$D{f}=""),"",'
           'ROUND(MOD($D{f}-$C{f},24)-IF($F{f}="",0,$F{f}),2))'.format(f=fila))
        ws.cell(row=fila, column=5).number_format = motor.FMT_DEC2
        ws.cell(row=fila, column=5).alignment = CENTRO
    motor.marcar_verde(ws, 'B5:D12')
    motor.marcar_verde(ws, 'F5:F12')
    _desverdear(ws, 'E5:E12')

    ws['B13'] = ('El PARTIDO son 10:00-15:00 y 19:00-23:00: 13 h entre '
                 'extremos y 9 h efectivas, porque lleva 4 h de PAUSA. La '
                 'columna «Horas» ya NO se teclea: sale de fin − inicio menos '
                 'la pausa, así que cambiar un horario recalcula de verdad el '
                 'cuadrante y las alertas. El inicio y el fin son además los '
                 'que miden el descanso entre jornadas.')
    ws['B14'] = ('El DOBLE (D) es el turno de 16 h que dispara la alerta de '
                 'jornada diaria. Existe para poder MARCARLO y que salte, no '
                 'para planificarlo.')
    ws['B15'] = ('Escribe las horas como número: 7 son las 07:00 y 15,5 las '
                 '15:30. La columna A no se edita: esas 8 letras son las del '
                 'desplegable del cuadrante y las del calendario del 05.')
    for r in (13, 14, 15):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True,
                                                       vertical='top')
        ws.row_dimensions[r].height = 32
        ws.merge_cells('B{0}:F{0}'.format(r))

    cambios.append('{}:Turnos!A5:F12: tabla de los 8 turnos + parámetros B2 '
                   '(jornada diaria máx. {} h) y B3 (descanso mínimo {} h, '
                   'art. 34.3 ET). La columna «Horas» (E) pasa de constante '
                   'verde a FÓRMULA derivada de la hora de inicio, la de fin '
                   'y la PAUSA nueva (F): antes cambiar el horario de un '
                   'turno dejaba intactos los totales del cuadrante y la '
                   'alerta de jornada diaria, en contra de lo que prometían '
                   'las Instrucciones — §2/RT-02'
                   .format(F01, motor.PARAMETROS['jornada_diaria_max'][1],
                           motor.PARAMETROS['descanso_min'][1]))
    return ws


# ==========================================================================
# 01 — cuadrante semanal: bloque auxiliar + las 4 alertas
# ==========================================================================
def _aux_formulas(fila):
    """Las 27 fórmulas del bloque auxiliar de una fila (§2).

    ⚠ TODAS devuelven la cadena vacía cuando el día no tiene turno, y las
    alertas que las leen neutralizan ese vacío con `IF(x="",neutro,x)` en vez
    de comparar contra él. No es estilo: es lo único que hace MEDIBLE la
    batería §5 de `main.py`, y se descubrió midiendo, el 2026-08-24:

      · **pycel devuelve el valor CACHEADO de una celda de fórmula si el libro
        lo trae**, y `main.py` corre `inject_cache` (paso 3) ANTES de las
        demostraciones (paso 7). Con el auxiliar devolviendo `0`/`24` en la
        hoja vacía, `inject_cache` graba esos ceros y `demo_descanso` —que
        planifica `B6='T'`, `C6='M'` y lee `K6`— seguía leyendo el 24 cacheado:
        el gate daba «sin alerta» con la alerta perfectamente escrita. Medido:
        con cache, `P6` vale 0,0 con `B6='M'`; sobre el MISMO libro re-guardado
        sin cache, vale 8.
      · Como la cadena vacía no se cachea (`inject_cache` la salta), la cadena
        `B6 → P6 → AK6 → K6` queda entera sin cache y pycel la recalcula. El
        efecto secundario es además el que pide §1.5: la hoja recién descargada
        enseña las columnas auxiliares y el total EN BLANCO, no una pared de
        ceros.
      · Y el mismo motivo obliga a que ninguna alerta lea un RANGO de celdas de
        fórmula: pycel tampoco propaga `set_value` a través de un nodo rango
        (`K6=IF(COUNTIF($AK6:$AP6,…))` nunca recalculaba). Por eso las seis
        transiciones se comparan una a una.

    Ojo con el 0: un día marcado L, V o B SÍ tiene horas (0), así que la
    transición tiene que excluirlo explícitamente o un día libre seguido de una
    mañana daría «7 h de descanso» y una alerta falsa.
    """
    fuera = []
    for j in range(7):
        dia = get_column_letter(2 + j)
        for base, col_tabla in ((AUX_HORAS, 5), (AUX_INI, 3), (AUX_FIN, 4)):
            coord = '{}{}'.format(get_column_letter(base + j), fila)
            fuera.append((coord,
                          '=IF(${d}{f}="","",'
                          'IFERROR(VLOOKUP(${d}{f},{t},{c},FALSE),0))'
                          .format(d=dia, f=fila, t=TURNOS_TABLA, c=col_tabla)))
    for j in range(6):
        h1 = get_column_letter(AUX_HORAS + j)
        h2 = get_column_letter(AUX_HORAS + j + 1)
        ini1 = get_column_letter(AUX_INI + j)
        ini2 = get_column_letter(AUX_INI + j + 1)
        fin1 = get_column_letter(AUX_FIN + j)
        coord = '{}{}'.format(get_column_letter(AUX_TRANS + j), fila)
        fuera.append((coord,
                      '=IF(OR(${h1}{f}="",${h2}{f}="",${h1}{f}=0,${h2}{f}=0),'
                      '"",IF(${fin1}{f}<=${ini1}{f},${ini2}{f}-${fin1}{f},'
                      '24-${fin1}{f}+${ini2}{f}))'
                      .format(h1=h1, h2=h2, ini1=ini1, ini2=ini2, fin1=fin1,
                              f=fila)))
    return fuera


def _alertas_semanal(fila):
    """Las CUATRO alertas que la landing vende desde el día uno y que el
    fichero no tenía (COM-01). Cada una cita su artículo.

    Sin guarda de `$A6=""`: no hace falta y estorba. No hace falta porque una
    fila recién descargada deja el bloque auxiliar en blanco y las cuatro
    salen vacías por su propia aritmética (§1.5, y así lo mide el gate). Y
    estorba porque condicionaría la alerta a haber TECLEADO EL NOMBRE: quien
    planifica reparte turnos antes de escribir nada en la columna A, que es
    justo cuando la alerta tiene que estar viva — y es lo que hace la batería
    §5 de `main.py`, que planifica `B6`/`C6` y no toca `A6`.
    """
    p = [get_column_letter(AUX_HORAS + j) for j in range(7)]
    tr = [get_column_letter(AUX_TRANS + j) for j in range(6)]
    d, m = TURNOS_DESCANSO, TURNOS_JORNADA_MAX
    dias = '+'.join('IF(${c}{f}="",0,IF(${c}{f}>0,1,0))'.format(c=c, f=fila)
                    for c in p)
    return [
        # el total se calla con la fila vacía en vez de enseñar 30 ceros
        ('I{}'.format(fila),
         '=IF(COUNT(${a}{f}:${b}{f})=0,"",ROUND(SUM(${a}{f}:${b}{f}),2))'
         .format(a=p[0], b=p[-1], f=fila)),
        # «neutro = el propio umbral»: sin encadenamiento, d<d es FALSO sea
        # cual sea el convenio. Un literal (24, 99) mentiría con un umbral alto.
        ('K{}'.format(fila),
         '=IF(OR({o}),"⛔ <"&{d}&" h","")'
         .format(o=','.join('IF(${c}{f}="",{d},${c}{f})<{d}'
                            .format(c=c, f=fila, d=d) for c in tr), d=d)),
        ('L{}'.format(fila),
         '=IF(7-({c})=0,"⛔ 0 días",IF(7-({c})=1,'
         '"⚠ 1 día (el ET pide 1,5, acumulable en 14 días — art. 37.1)",""))'
         .format(c=dias)),
        ('M{}'.format(fila),
         '=IF(OR($I{f}="",$J{f}=""),"",IF($I{f}>$J{f},'
         '"⛔ +"&TEXT($I{f}-$J{f},"0")&" h",""))'.format(f=fila)),
        # RD-16 · N usa 8 h en vez del umbral del convenio cuando la fila es
        # de un menor: para un menor no hay distribución irregular que suba la
        # jornada ordinaria (art. 34.3 ET por remisión del art. 6).
        ('N{}'.format(fila),
         '=IF(${men}{f}="S",IF(OR({om}),"⛔ jornada > {jm} h (MENOR)",""),'
         'IF(OR({o}),"⛔ jornada > "&{m}&" h",""))'
         .format(men=COL_MENOR, f=fila, jm=JORNADA_MENOR, m=m,
                 om=','.join('IF(${c}{f}="",0,${c}{f})>{jm}'
                             .format(c=c, f=fila, jm=JORNADA_MENOR)
                             for c in p),
                 o=','.join('IF(${c}{f}="",0,${c}{f})>{m}'
                            .format(c=c, f=fila, m=m) for c in p))),
        # RD-16 · la alerta que no existía: nocturnidad y turno doble
        # prohibidos al menor de 18 años (art. 6.2 y 6.3 ET). Se comparan los
        # SIETE días uno a uno, no con un COUNTIF sobre el rango: pycel no
        # propaga `set_value` a través de un nodo rango (ver `_aux_formulas`).
        ('{}{}'.format(COL_ALERTA_MENOR, fila),
         '=IF(${men}{f}<>"S","",IF(OR({no}),'
         '"⛔ MENOR: trabajo nocturno y turno doble prohibidos (art. 6 ET)",'
         '"⚠ menor: sin horas extra (art. 6.3 ET)"))'
         .format(men=COL_MENOR, f=fila,
                 no=','.join('${c}{f}="N",${c}{f}="D"'
                             .format(c=get_column_letter(2 + j), f=fila)
                             for j in range(7)))),
    ]


def _cuadrante_semanal(ws, cambios):
    r0, r1_hoy, r1 = 6, 20, 5 + CAP
    delta = motor.expandir_filas(ws, r1_hoy, r1, cola=_cola(ws, r1_hoy))
    if delta:
        cambios.append('{}:Cuadrante Semanal: bloque 6..{} → 6..{} '
                       '({} empleados, §1.3) y la cola bajada {} filas'
                       .format(F01, r1_hoy, r1, CAP, delta))

    _cab(ws, 5, ['Empleado'] + list(DIAS)
         + ['Total Horas', 'H. contratadas/semana',
            'Descanso entre jornadas', 'Descanso semanal', 'Jornada semanal',
            'Jornada diaria', 'Menor de edad (S/N)', 'Alerta menores'])
    for letra, ancho in (('I', 13), ('J', 15), ('K', 20), ('L', 26),
                         ('M', 18), ('N', 22), ('O', 18), ('P', 34)):
        ws.column_dimensions[letra].width = ancho

    # etiquetas del bloque auxiliar en la fila 4 (el motor sólo mira la 5)
    for col, texto in ((AUX_HORAS, 'AUXILIAR — no editar · horas por día'),
                       (AUX_INI, 'hora de inicio'),
                       (AUX_FIN, 'hora de fin'),
                       (AUX_TRANS, 'descanso entre jornadas (h)')):
        cel = ws.cell(row=4, column=col, value=texto)
        cel.font = Font(italic=True, size=8, color='999999')

    n = 0
    for fila in range(r0, r1 + 1):
        for coord, formula in _aux_formulas(fila) + _alertas_semanal(fila):
            _f(ws, coord, formula)
            n += 1
        cel = ws.cell(row=fila, column=10)         # J · h. contratadas
        cel.value = motor.PARAMETROS['jornada_semanal'][1]
        cel.number_format = motor.FMT_DEC2
        cel.alignment = CENTRO
        # RD-16 · O es la ÚNICA forma de cumplir sin enlazar libros (§6): el
        # usuario copia la S del aviso del 07 y el 01 empieza a mirar los
        # límites del menor.
        cel = ws.cell(row=fila, column=15)         # O · menor de edad
        cel.alignment = CENTRO
        ws.cell(row=fila, column=16).alignment = Alignment(  # P · alerta
            wrap_text=True, vertical='center')
    motor.marcar_verde(ws, '{c}{a}:{c}{b}'.format(c=COL_MENOR, a=r0, b=r1))
    _limpiar_mis_dv(ws, ('{c}{a}:{c}{b}'.format(c=COL_MENOR, a=r0, b=r1),))
    _dv(ws, '{c}{a}:{c}{b}'.format(c=COL_MENOR, a=r0, b=r1), ['S', 'N'],
        'Valor no válido',
        'Escribe S si esa persona es MENOR de 18 años. Lo sabe el fichero 07: '
        'su columna «Aviso» te lo dice en cuanto pones la fecha de '
        'nacimiento. Con una S, la alerta de jornada diaria baja a '
        '{jm} h y salta si le pones un turno de noche (N) o un doble (D): '
        'art. 6 ET.'.format(jm=JORNADA_MENOR),
        'Escribe S o N.')
    motor._limpiar_cf(ws, set(['{c}{a}:{c}{b}'
                               .format(c=COL_ALERTA_MENOR, a=r0, b=r1)]))
    motor.semaforo(ws, '{c}{a}:{c}{b}'.format(c=COL_ALERTA_MENOR, a=r0, b=r1),
                   motor.VOC_ALERTA)
    for col in range(AUX_HORAS, AUX_ULT + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # el TOTAL del equipo lo estira `expandir_filas`; la fila de PROMEDIO es
    # nueva y va justo debajo, con la guarda de rango vacío del §1.5.
    fila_total = _fila_con(ws, 1, 'TOTAL HORAS EQUIPO', r1 + 1)
    if fila_total:
        ws.cell(row=fila_total, column=11).value = None
        # RD-30 · las columnas de ALERTA sólo devuelven texto y arrastraban el
        # '0.00' del bloque numérico vecino (L, que hace lo mismo, iba en
        # 'General'). Lo mismo las dos cabeceras nuevas.
        for col in ('K', 'M', 'N', COL_MENOR, COL_ALERTA_MENOR):
            for f in range(r0, r1 + 1):
                if col != COL_MENOR:
                    ws['{}{}'.format(col, f)].number_format = 'General'
        for col in ('I', 'J', 'K', 'L', 'M', 'N', COL_MENOR,
                    COL_ALERTA_MENOR):
            ws['{}5'.format(col)].number_format = 'General'
        # RD-30 · el único total sin guarda de rango vacío era precisamente
        # el que se ve al abrir el fichero: `I36 = '=SUM(I6:I35)'`, con cache
        # 0,0, mientras su vecina K36 sí la llevaba. La convención de familia
        # es que ninguna hoja recién descargada enseñe un contador (§1.5).
        _f(ws, 'I{}'.format(fila_total),
           '=IF(COUNT($I${a}:$I${b})=0,"",ROUND(SUM($I${a}:$I${b}),2))'
           .format(a=r0, b=r1))
        ws['I{}'.format(fila_total)].number_format = motor.FMT_DEC2
        _f(ws, 'K{}'.format(fila_total),
           motor.guarda_media('$I${}:$I${}'.format(r0, r1)))
        ws.cell(row=fila_total, column=10,
                value='Media por empleado →').font = NEGRITA
        ws.cell(row=fila_total, column=10).alignment = Alignment(
            horizontal='right')
        n += 1

    cambios.append('{}:Cuadrante Semanal!O{}:P{}: casilla verde «Menor de '
                   'edad (S/N)» (se copia del aviso del 07) + alerta propia: '
                   'con S, la jornada diaria se mide contra {} h y salta si '
                   'aparece un turno N (noche) o D (doble) — art. 6 ET. Era '
                   'la única infracción del kit que se cometía en el 01 y sólo '
                   'se avisaba en el 07 (RD-16)'
                   .format(F01, r0, r1, JORNADA_MENOR))
    cambios.append('{}:Cuadrante Semanal!I{}:N{}: {} fórmulas — bloque '
                   'auxiliar oculto {}:{} (horas, inicio, fin y las 6 '
                   'transiciones) y las 4 alertas legales: descanso entre '
                   'jornadas (art. 34.3 ET, {} h), descanso semanal '
                   '(art. 37.1), jornada semanal contra las horas '
                   'CONTRATADAS de cada uno (J, verde, {} por defecto) y '
                   'jornada diaria / turno doble (art. 34.3, {} h) — '
                   'DOM-05/DOM-25/TEC-05/COM-01'
                   .format(F01, r0, r1, n, get_column_letter(AUX_HORAS),
                           get_column_letter(AUX_ULT),
                           motor.PARAMETROS['descanso_min'][1],
                           motor.PARAMETROS['jornada_semanal'][1],
                           motor.PARAMETROS['jornada_diaria_max'][1]))
    return r0, r1


# ==========================================================================
# 01 — `Cuadrante Mensual` (DOM-17/TEC-23/COM-12)
# ==========================================================================
def _horas_mes(fila):
    """Horas de la semana en la hoja mensual. El bloque auxiliar del semanal
    no se puede replicar cinco veces (serían 135 columnas), así que los siete
    `VLOOKUP` van inline: es la MISMA tabla `Turnos`, así que las dos hojas no
    pueden discrepar."""
    trozos = []
    for j in range(7):
        trozos.append('IFERROR(VLOOKUP(${d}{f},{t},5,FALSE),0)'
                      .format(d=get_column_letter(2 + j), f=fila,
                              t=TURNOS_TABLA))
    return ('=IF($A{f}="","",ROUND({s},2))'
            .format(f=fila, s='+'.join(trozos)))


def _cuadrante_mensual(wb, cambios, r0_sem):
    """Cinco semanas VIVAS + el total del mes por empleado.

    Se borra y se vuelve a crear entera: hoy son 30 rótulos y CERO fórmulas,
    CERO validación y CERO alertas para 28 días (`01-cuadrante-turnos-semanal
    .xlsx:Cuadrante Mensual:A3..A61`), así que no hay nada que conservar y una
    reconstrucción completa es lo único idempotente de verdad.

    Los nombres NO se teclean: salen del cuadrante semanal (`=IF('Cuadrante
    Semanal'!$A6="","",…)`), que es lo que evita mantener la plantilla en cinco
    sitios. La fila 3 la ocupa la leyenda única que escribe el motor (§1.1) y
    por eso «SEMANA 1» va en `A3`: la fila 4 tiene que ser la de CABECERA,
    que es la que `motor.PRESENTACION` inmoviliza (`A5`) y repite al imprimir
    (`$4:$4`) — y esa cabecera es idéntica en los cinco bloques.
    """
    idx = wb.sheetnames.index('Cuadrante Mensual') \
        if 'Cuadrante Mensual' in wb.sheetnames else len(wb.sheetnames)
    if 'Cuadrante Mensual' in wb.sheetnames:
        wb.remove(wb['Cuadrante Mensual'])
    ws = wb.create_sheet('Cuadrante Mensual', idx)

    ws.column_dimensions['A'].width = 22
    for j in range(7):
        ws.column_dimensions[get_column_letter(2 + j)].width = 12
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 26

    ws.merge_cells('A1:J1')
    ws['A1'] = 'Cuadrante Mensual — Mes: _______________'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A2:J2')
    ws['A2'] = 'AI Chef Pro · aichef.pro — Kit Gestión de Personal y Turnos'
    ws['A2'].font = Font(size=9, color='666666')

    cabecera = ['Empleado'] + list(DIAS) + ['Total Horas', 'Alerta semanal']
    semanas, n = [], 0
    fila = 3
    for semana in range(1, 6):
        if semana == 1:
            ws['A3'] = 'SEMANA 1'
            ws['A3'].font = Font(bold=True, size=11)
            hdr = 4
        else:
            _titulo_bloque(ws, fila, 'SEMANA {}'.format(semana), 10)
            hdr = fila + 1
        _cab(ws, hdr, cabecera)
        d0 = hdr + 1
        d1 = hdr + CAP
        semanas.append((d0, d1))
        for k in range(CAP):
            f = d0 + k
            sem = r0_sem + k                     # fila gemela del semanal
            _f(ws, 'A{}'.format(f),
               '=IF(\'Cuadrante Semanal\'!$A{s}="","",'
               '\'Cuadrante Semanal\'!$A{s})'.format(s=sem))
            _f(ws, 'I{}'.format(f), _horas_mes(f))
            _f(ws, 'J{}'.format(f),
               '=IF($A{f}="","",IF($I{f}>\'Cuadrante Semanal\'!$J{s},'
               '"⛔ +"&TEXT($I{f}-\'Cuadrante Semanal\'!$J{s},"0")&" h",""))'
               .format(f=f, s=sem))
            n += 3
            for j in range(7):
                cel = ws.cell(row=f, column=2 + j)
                cel.alignment = CENTRO
        motor.marcar_verde(ws, 'B{}:H{}'.format(d0, d1))
        _dv(ws, 'B{}:H{}'.format(d0, d1),
            motor.DV_JORNADA.split(','), 'Código de turno no válido',
            motor.LEYENDA_JORNADA + '.',
            'Usa uno de los 8 códigos del kit: son los mismos que en el '
            '«Cuadrante Semanal» y de ellos salen las horas y las alertas.')
        motor._limpiar_cf(ws, set(['J{}:J{}'.format(d0, d1)]))
        motor.semaforo(ws, 'J{}:J{}'.format(d0, d1), motor.VOC_ALERTA)
        fila = d1 + 2

    # ---- total del mes por empleado + fila de promedio -------------------
    _titulo_bloque(ws, fila, 'TOTAL DEL MES POR EMPLEADO', 10)
    hdr = fila + 1
    _cab(ws, hdr, ['Empleado', 'Semana 1', 'Semana 2', 'Semana 3', 'Semana 4',
                   'Semana 5', 'Total del mes', 'Media semanal',
                   'H. contratadas/semana', 'Alerta del cómputo'])
    t0, t1 = hdr + 1, hdr + CAP
    for k in range(CAP):
        f = t0 + k
        sem = r0_sem + k
        _f(ws, 'A{}'.format(f),
           '=IF(\'Cuadrante Semanal\'!$A{s}="","",'
           '\'Cuadrante Semanal\'!$A{s})'.format(s=sem))
        for w, par in enumerate(semanas):
            _f(ws, '{}{}'.format(get_column_letter(2 + w), f),
               '=IF($A{f}="","",$I{g})'.format(f=f, g=par[0] + k))
        _f(ws, 'G{}'.format(f),
           '=IF($A{f}="","",ROUND(SUM($B{f}:$F{f}),2))'.format(f=f))
        _f(ws, 'H{}'.format(f),
           motor.guarda_media('$B{f}:$F{f}'.format(f=f)))
        _f(ws, 'I{}'.format(f),
           '=IF(\'Cuadrante Semanal\'!$A{s}="","",'
           '\'Cuadrante Semanal\'!$J{s})'.format(s=sem))
        _f(ws, 'J{}'.format(f),
           '=IF(OR($H{f}="",$I{f}=""),"",IF($H{f}>$I{f},'
           '"⛔ media +"&TEXT($H{f}-$I{f},"0")&" h/semana",""))'.format(f=f))
        n += 5 + len(semanas)
    motor._limpiar_cf(ws, set(['J{}:J{}'.format(t0, t1)]))
    motor.semaforo(ws, 'J{}:J{}'.format(t0, t1), motor.VOC_ALERTA)

    prom = t1 + 1
    ws.cell(row=prom, column=1, value='PROMEDIO DEL EQUIPO').font = NEGRITA
    for col in ('B', 'C', 'D', 'E', 'F', 'G', 'H'):
        _f(ws, '{}{}'.format(col, prom),
           motor.guarda_media('${c}${a}:${c}${b}'.format(c=col, a=t0, b=t1)))
        n += 1
    for c in range(1, 11):
        ws.cell(row=prom, column=c).fill = PatternFill('solid',
                                                       fgColor='EEEEEE')
        ws.cell(row=prom, column=c).font = NEGRITA

    # RT-17 · un salto de página ANTES de cada bloque: con el título del
    # documento repetido arriba ('$1:$2', motor.PRESENTACION) y un bloque por
    # página, cada semana se imprime con su rótulo «SEMANA n» y su cabecera.
    ws.row_breaks.brk = []
    for d0, _d1 in semanas[1:]:
        ws.row_breaks.append(Break(id=d0 - 3))
    ws.row_breaks.append(Break(id=hdr - 2))

    pie = prom + 2
    ws.merge_cells('A{0}:J{0}'.format(pie))
    ws.cell(row=pie, column=1, value=motor.PIE).font = Font(size=9,
                                                            color='666666')
    # RD-30 · la columna J («Alerta semanal» / «Alerta del cómputo») sólo
    # devuelve texto y llevaba '0.00', igual que las dos CABECERAS de la fila
    # 4. Un formato numérico sobre una cabecera de texto es el resto de
    # formato que hace dudar de la hoja al cliente que la imprime.
    for f in range(4, prom + 1):
        cabecera = ws.cell(row=f, column=1).value == 'Empleado'
        ws.cell(row=f, column=9).number_format = (
            'General' if cabecera else motor.FMT_DEC2)
        ws.cell(row=f, column=10).number_format = 'General'

    cambios.append('{}:Cuadrante Mensual: hoja reconstruida — 5 semanas × {} '
                   'empleados con DV de los 8 códigos, horas por VLOOKUP a '
                   '«Turnos», alerta semanal contra las horas contratadas, '
                   'nombres ENLAZADOS al cuadrante semanal, total del mes por '
                   'empleado (G{}:G{}), media semanal y fila de promedio '
                   '({} fórmulas). Antes: 30 rótulos y cero fórmulas para 28 '
                   'días — DOM-17/TEC-23/COM-12'
                   .format(F01, CAP, t0, t1, n))
    return ws


def _instrucciones_01(wb, cambios):
    ley = _lineas_leyenda(F01)
    lineas = [
        'Cómo usar esta plantilla:',
        '▸ Escribe los nombres de tu equipo UNA sola vez, en la columna A de '
        'esta hoja: el \'Cuadrante Mensual\' los arrastra solo y no hay que '
        'teclearlos cinco veces.',
        ley[0] if ley else None,
        '▸ Las horas NO están escritas dentro de la fórmula: salen de la hoja '
        '\'Turnos\'. Si tu mañana empieza a las 8:00, cámbialo allí y se '
        'recalcula el kit entero: la columna «Horas» de esa hoja es una '
        'fórmula (fin − inicio − pausa), no un número que haya que corregir a '
        'mano.',
        ley[1] if len(ley) > 1 else None,
        '▸ La hoja \'Cuadrante Mensual\' repite la rejilla cinco semanas '
        '(los meses de 31 días no caben en cuatro), hereda los nombres del '
        '\'Cuadrante Semanal\', suma el total del mes por empleado y saca la '
        'media semanal, que es lo que se compara con el contrato cuando '
        'distribuyes la jornada de forma irregular.',
        None,
        'Las 4 alertas del cuadrante (columnas K a N):',
        '▸ K · Descanso entre jornadas: compara la hora de FIN de cada día con '
        'la de INICIO del siguiente, contando el turno que cruza la '
        'medianoche. Salta por debajo de las 12 h del art. 34.3 ET; el umbral '
        'es la celda Turnos!B3.',
        '▸ L · Descanso semanal: art. 37.1 ET, día y medio ininterrumpido, '
        'acumulable en periodos de hasta 14 días. Con cero días libres en la '
        'semana la alerta es roja; con uno, ámbar.',
        '▸ M · Jornada semanal: compara con las horas CONTRATADAS de esa '
        'persona (columna J, verde, {jor} por defecto), no con un 40 fijo: a '
        'un contrato de 20 h las alertas le saltan a las 20.',
        '▸ N · Jornada diaria: art. 34.3 ET, máximo {max} h ordinarias al día '
        'salvo distribución irregular pactada (umbral en Turnos!B2). Es la que '
        'caza el turno doble.',
        '▸ O y P · Menores de edad: marca S en la columna O (te lo dice el '
        'aviso del fichero 07 en cuanto pones la fecha de nacimiento). Con la '
        'S, la alerta N baja a 8 h —a un menor no le vale la distribución '
        'irregular— y la P salta si le has puesto un turno de NOCHE (N) o un '
        'DOBLE (D): al menor de 18 años le están prohibidos el trabajo '
        'nocturno y las horas extraordinarias (art. 6 ET).',
        None,
        'La tabla de turnos vive en la hoja \'Turnos\':',
        '▸ Cada código lleva su hora de inicio, su hora de fin y su PAUSA; '
        'las horas efectivas se calculan solas. Las horas se escriben como '
        'número: 7 = 07:00 y 15,5 = 15:30.',
        '▸ El PARTIDO (P) son 9 h efectivas entre las 10:00 y las 23:00 '
        '(13 h entre extremos menos 4 de pausa); el DOBLE (D) son 16 h y '
        'existe para que la alerta de jornada diaria tenga algo que cazar.',
        '▸ V (vacaciones) y B (baja) se marcan aquí con las mismas letras que '
        'en el 05 y cuentan 0 h. El permiso es PE y sólo existe en el 05: en '
        'este cuadrante la P es el turno Partido.',
        '▸ Una letra que no esté en la tabla cuenta 0 h: por eso las celdas de '
        'turno llevan desplegable.',
        '▸ Las columnas R a AR están OCULTAS: son el cálculo auxiliar (horas, '
        'hora de inicio, hora de fin y las 6 transiciones entre jornadas). '
        'Muéstralas si quieres '
        'auditar una alerta.',
    ]
    lineas = [x.format(jor=motor.PARAMETROS['jornada_semanal'][1],
                       max=motor.PARAMETROS['jornada_diaria_max'][1])
              if isinstance(x, str) and '{' in x else x for x in lineas]
    _instrucciones(wb['Instrucciones'], lineas)
    cambios.append('{}:Instrucciones: cuerpo reescrito — las 4 alertas '
                   'descritas una a una con su artículo, el descanso pasa a '
                   '12 h (art. 34.3 ET) y aparecen por fin la hoja '
                   '\'Turnos\' y el \'Cuadrante Mensual\', hoy no '
                   'mencionado ni una vez (TEC-23)'.format(F01))


# ==========================================================================
# 02 — registro de horas
# ==========================================================================
CAB_REGISTRO = ['Empleado', 'Fecha', 'Entrada', 'Salida', 'Pausa (h)',
                'Horas trabajadas', 'H. contratadas', 'Horas extra', 'Tipo',
                'Aviso']
COL_AVISO_REG = 'J'

#: RD-18 · «Complementaria (contrato parcial)» es una figura DISTINTA de la
#: hora extraordinaria y el kit no la tenía: al contrato a tiempo parcial le
#: están PROHIBIDAS las horas extraordinarias (art. 12.4.c ET) y sus horas de
#: más son complementarias — exigen pacto escrito, tienen su propio tope y se
#: retribuyen COMO ORDINARIAS (art. 12.5 ET). Aplicarles el 1,25 y contarlas
#: contra las 80 h/año del art. 35.2 es un error que recoge un acta de
#: inspección, y afecta a media plantilla de hostelería.
TIPO_COMPLEMENTARIA = 'Complementaria (contrato parcial)'
TIPO_COMPENSADA = 'Compensada con descanso'
TIPOS_EXTRA = ['Voluntaria', 'Obligatoria', 'Fuerza mayor',
               TIPO_COMPENSADA, TIPO_COMPLEMENTARIA]


def _pre_02(wb, cambios):
    """La columna «Pausa (h)» se inserta en `pre()`, ANTES de que el motor fije
    rangos (regla del pipeline). Es lo que hace registrable el turno partido,
    que hoy no cabe en un solo par entrada/salida: 10:00→23:00 da 13 h y 5 h
    extra falsas, y partido en dos filas contra 8 h contratadas cada una da
    CERO extras cuando la realidad es 1 h (DOM-21)."""
    ws = wb['Registro Horas']
    if isinstance(ws['E4'].value, str) and 'pausa' in ws['E4'].value.lower():
        return 0                                   # ya insertada (2.ª pasada)
    motor.insertar_columna(ws, 5)
    cambios.append('{}:Registro Horas!E: columna «Pausa (h)» insertada; el '
                   'Tipo de hora extra se desplaza a la I — DOM-21'
                   .format(F02))
    return 1


#: Tramo entrada→salida en horas, con el cruce de medianoche del §7-bis.2.
_TRAMO = 'MOD($D{f}-$C{f},1)*24'
_PAUSA = 'IF($E{f}="",0,$E{f})'


def _horas_registro(fila):
    """RD-19/RT-09/RT-15 — «Horas trabajadas» con las tres guardas que le
    faltaban, sin dejar de ser un NÚMERO (la columna H la resta).

    Tres formas medidas de perder una jornada entera en silencio:

      · **hora tecleada como número** (`C=9`, `D=17`): `MOD(8,1)` es 0 y la
        celda muestra `00:00`, que se lee como «aún no lo he puesto». El
        `hh:mm` de la v2.0 cambió el síntoma —antes daban 192 h— y no el
        agujero.
      · **pausa mayor que el tramo** (`09:00→13:00` con `E=6`): devolvía
        `−2,00 h`, un asiento inválido en el documento del art. 34.9 ET. El
        `MAX(0,…)` de la columna H lo tapaba aguas abajo, así que no saltaba
        en ningún total.
      · **entrada igual que salida**: `MOD(0,1)` es 0, indistinguible de un día
        no trabajado. Es el error de transcripción típico de un parte de
        firmas.

    En los tres casos la celda se queda VACÍA —no en un cero que parece un
    dato— y la columna «Aviso» dice cuál de los tres es.
    """
    tramo, pausa = _TRAMO.format(f=fila), _PAUSA.format(f=fila)
    return ('=IF(OR($C{f}="",$D{f}=""),"",'
            'IF(OR($C{f}>=1,$D{f}>=1,$C{f}=$D{f},{p}>{t}),"",'
            'ROUND({t}-{p},2)))'.format(f=fila, t=tramo, p=pausa))


#: Bloque de empleados del «Resumen Mensual» del 02 (cabecera en la 5).
RES0, RES1 = 6, 5 + CAP


def _aviso_registro(fila, r0=RES0, r1=RES1):
    """RD-19/RT-05/RT-09/RT-15 — la columna que dice por qué una fila no suma.

    El cuarto caso es RT-05, y es el más caro: `Resumen Mensual!B6` agrega con
    `SUMIF` por NOMBRE, y el nombre se teclea DOS veces a mano, en dos hojas,
    sin lista y sin comprobación. Un espacio final, una tilde de menos o
    «Ana P.» frente a «Ana Pérez» y el resumen devuelve 0,00 h y 0,00 € para
    esa persona: las horas extra existen en el registro, están calculadas, y
    desaparecen del cómputo y del coste sin un solo aviso. En el 05 este mismo
    riesgo SÍ estaba cubierto (`Solicitudes!H5`, «⚠ ese nombre no está en el
    calendario»); en el 02 —el fichero del dinero y del registro obligatorio—
    no lo estaba.
    """
    tramo, pausa = _TRAMO.format(f=fila), _PAUSA.format(f=fila)
    return ('=IF(AND($A{f}="",$C{f}="",$D{f}=""),"",'
            'IF(OR($C{f}>=1,$D{f}>=1),'
            '"⚠ escribe la hora con dos puntos: 9:00, no 9",'
            'IF(AND($C{f}<>"",$C{f}=$D{f}),'
            '"⚠ entrada y salida iguales: la jornada no suma",'
            'IF(AND($C{f}<>"",$D{f}<>"",{p}>{t}),'
            '"⚠ la pausa es mayor que la jornada",'
            'IF(AND($A{f}<>"",'
            'COUNTIF(\'Resumen Mensual\'!$A${a}:$A${b},$A{f}&"")=0),'
            '"⚠ ese nombre no está en el «Resumen Mensual»: sus horas no se '
            'agregan","")))))'
            .format(f=fila, t=tramo, p=pausa, a=r0, b=r1))


def _registro_horas(ws, cambios):
    r0, r1_hoy, r1 = 5, 54, 4 + FILAS_REG
    delta = motor.expandir_filas(ws, r1_hoy, r1, cola=_cola(ws, r1_hoy))
    if delta:
        cambios.append('{}:Registro Horas: {} → {} filas de registro '
                       '(§1.3/COM-18)'.format(F02, r1_hoy - r0 + 1, FILAS_REG))
    _cab(ws, 4, CAB_REGISTRO)
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['I'].width = 26
    ws.column_dimensions[COL_AVISO_REG].width = 40

    for fila in range(r0, r1 + 1):
        _f(ws, 'F{}'.format(fila), _horas_registro(fila))
        _f(ws, 'H{}'.format(fila),
           motor.guarda_resta('$F{}'.format(fila), '$G{}'.format(fila),
                              'MAX(0,$F{f}-$G{f})'.format(f=fila)))
        _f(ws, '{}{}'.format(COL_AVISO_REG, fila), _aviso_registro(fila))
        ws['{}{}'.format(COL_AVISO_REG, fila)].alignment = Alignment(
            wrap_text=True, vertical='center')
        cel = ws.cell(row=fila, column=7)          # G · H. contratadas
        cel.value = 8
        cel.number_format = motor.FMT_DEC2

    ref_av = '{c}{a}:{c}{b}'.format(c=COL_AVISO_REG, a=r0, b=r1)
    motor._limpiar_cf(ws, set([ref_av]))
    motor.semaforo(ws, ref_av, motor.VOC_ALERTA)
    _limpiar_mis_dv(ws, ('C{}:D{}'.format(r0, r1),
                         'E{}:E{}'.format(r0, r1),
                         'G{}:G{}'.format(r0, r1), 'I{}:I{}'.format(r0, r1)))
    # RD-19 · el formato `hh:mm` cambió el síntoma y no cerró el agujero:
    # teclear «9» y «17» ya no daba 192 h, daba 0,00 h — una jornada entera
    # desaparecía del registro sin un solo aviso, y el registro de jornada es
    # el documento que hay que conservar 4 años (art. 34.9 ET). La DV lo
    # RECHAZA en la entrada y el aviso de la columna J lo explica si alguien
    # pega valores (el pegado se salta la validación).
    _dv(ws, 'C{}:D{}'.format(r0, r1), None, 'Hora no válida',
        'Escribe la hora con dos puntos: 9:00, no 9. Un 9 suelto es el '
        'NÚMERO nueve (nueve días), no las nueve de la mañana, y la columna '
        '«Horas trabajadas» se quedaría en 0,00 sin avisar.',
        'Escribe una hora entre 0:00 y 23:59, con dos puntos (9:00).',
        tipo='time', operator='between', formula1='0', formula2='0.9993')
    # RT-09 · la pausa no tenía NINGUNA validación: teclear 6 en una jornada
    # de 4 h devolvía −2,00 h en un asiento del registro de jornada.
    _dv(ws, 'E{}:E{}'.format(r0, r1), None, 'Pausa no válida',
        'Pausa NO trabajada de ese día, en horas: 0,5 son treinta minutos. '
        'El turno partido de 10:00 a 23:00 lleva 4. Si la pausa sale mayor '
        'que la jornada, la columna «Horas trabajadas» se queda en blanco y '
        'la de «Aviso» te dice por qué.',
        'Escribe un número de horas entre 0 y 8.',
        tipo='decimal', operator='between', formula1='0', formula2='8')
    _dv(ws, 'G{}:G{}'.format(r0, r1), None, 'Jornada diaria no válida',
        'Jornada CONTRATADA de ese día, en horas. Viene precargada a 8. Si la '
        'dejas vacía, la columna «Horas extra» se queda en blanco a propósito: '
        'sin saber lo contratado no se puede decir qué es extra.',
        'Escribe un número de horas entre 1 y 12.',
        tipo='decimal', operator='between', formula1='1', formula2='12')
    _dv(ws, 'I{}:I{}'.format(r0, r1), TIPOS_EXTRA,
        'Tipo de hora extra no válido',
        'Art. 35.2 ET: las de FUERZA MAYOR y las COMPENSADAS CON DESCANSO '
        'dentro de los 4 meses siguientes NO computan en el tope de 80 h/año. '
        'El «Resumen Mensual» las descuenta buscando este texto exacto.',
        'Elige un valor de la lista: las fórmulas de agregación buscan este '
        'texto exacto.')
    cambios.append('{}:Registro Horas!F{}:I{}: horas con MOD (cruce de '
                   'medianoche: 23:00→07:00 son 8,00 h, no −16) menos la '
                   'pausa, y la hora extra con GUARDA de «H. contratadas» '
                   'vacía. G precargada a 8 con DV 1-12; el Tipo pasa a las 4 '
                   'modalidades del art. 35.2 — DOM-03/DOM-11/DOM-21/TEC-01/'
                   'TEC-09/COM-07'.format(F02, r0, r1))
    return r0, r1


def _resumen_mensual(ws, cambios, reg0, reg1):
    r0, r1_hoy, r1 = 6, 20, 5 + CAP
    delta = motor.expandir_filas(ws, r1_hoy, r1, cola=_cola(ws, r1_hoy))
    if delta:
        cambios.append('{}:Resumen Mensual: 15 → {} empleados (§1.3)'
                       .format(F02, CAP))

    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['E'].width = 20
    _limpiar_mis_dv(ws)
    _nota_parametro(ws, 3, 'tarifa_hora', 1, 2)
    _nota_parametro(ws, 3, 'recargo_extra', 3, 4)
    _nota_parametro(ws, 3, 'limite_extra_anual', 5, 6, alto=58)

    # ⚠ La columna E era «Coste ×1.75» y traía fórmula en E6:E20. Ahora es la
    # casilla VERDE del acumulado del año: si no se vacía, la hoja conserva un
    # coste calculado con el multiplicador que esta versión retira, `aplicar_
    # verde` la ve «calculada» y no la pinta, y el cliente no puede escribir en
    # ella. Lo mismo la F, que era «Coste ×2.0».
    for f in range(r0, r1 + 1):
        ws.cell(row=f, column=5).value = None
        ws.cell(row=f, column=5).number_format = motor.FMT_DEC2

    # RT-16/RC-20 · la cabecera de F se llamaba «Límite anual de horas extra»,
    # que es literalmente el rótulo del PARÁMETRO de E3/F3 dos filas más
    # arriba: anunciaba un número y devolvía «✓ dentro».
    _cab(ws, 5, ['Empleado', 'Total H. Extra del mes',
                 'H. extra no computables (art. 35.2 y 12.5)',
                 'H. extra computables', 'H. extra acumuladas en el año',
                 'Estado frente al límite anual',
                 'Coste de las horas extra (€)',
                 'Saldo del mes (trabajadas − contratadas)'])
    ws.column_dimensions['H'].width = 26

    reg = "'Registro Horas'"
    emp = '{r}!$A${a}:$A${b}'.format(r=reg, a=reg0, b=reg1)
    ext = '{r}!$H${a}:$H${b}'.format(r=reg, a=reg0, b=reg1)
    tip = '{r}!$I${a}:$I${b}'.format(r=reg, a=reg0, b=reg1)
    tra = '{r}!$F${a}:$F${b}'.format(r=reg, a=reg0, b=reg1)
    con = '{r}!$G${a}:$G${b}'.format(r=reg, a=reg0, b=reg1)

    def por_tipo(f, etiqueta):
        return ('SUMIFS({x},{e},$A{f}&"",{t},"{q}")'
                .format(x=ext, e=emp, t=tip, f=f, q=etiqueta))

    n = 0
    for f in range(r0, r1 + 1):
        _f(ws, 'B{}'.format(f),
           '=IF($A{f}="","",ROUND(SUMIF({e},$A{f}&"",{x}),2))'
           .format(f=f, e=emp, x=ext))
        # RD-18 · las complementarias del contrato parcial tampoco computan en
        # el tope del art. 35.2: no son horas extraordinarias, que al parcial
        # le están prohibidas (art. 12.4.c ET).
        _f(ws, 'C{}'.format(f),
           '=IF($A{f}="","",ROUND({fm}+{cd}+{co},2))'
           .format(f=f, fm=por_tipo(f, 'Fuerza mayor'),
                   cd=por_tipo(f, TIPO_COMPENSADA),
                   co=por_tipo(f, TIPO_COMPLEMENTARIA)))
        _f(ws, 'D{}'.format(f),
           '=IF($A{f}="","",ROUND($B{f}-$C{f},2))'.format(f=f))
        _f(ws, 'F{}'.format(f),
           '=IF($E{f}="","",IF($E{f}>$F$3,'
           '"⛔ EXCEDE ("&TEXT($E{f}-$F$3,"0")&" h)",'
           'IF($E{f}>$F$3*0.8,"⚠ cerca del límite","✓ dentro")))'.format(f=f))
        # RD-02/RT-06 · el coste se calculaba sobre B (TODAS las horas extra
        # del mes), compensadas con descanso incluidas. Esas horas, por
        # definición del art. 35.1 ET, se pagan con tiempo de descanso y NO con
        # dinero: el kit las pagaba dos veces (20 h compensadas a 12,00 € ×
        # 1,25 imputaban 300 € que nadie desembolsa, y encima ya se había dado
        # el descanso). Las de FUERZA MAYOR sí se pagan —sólo no computan en
        # el tope—, así que siguen dentro. Y las COMPLEMENTARIAS se retribuyen
        # como ORDINARIAS (art. 12.5 ET): entran sin el recargo (RD-18).
        _f(ws, 'G{}'.format(f),
           '=IF($B{f}="","",ROUND(($B{f}-{cd}-{co})*$B$3*$D$3'
           '+{co}*$B$3,2))'.format(f=f, cd=por_tipo(f, TIPO_COMPENSADA),
                                   co=por_tipo(f, TIPO_COMPLEMENTARIA)))
        # RD-17 · la hora extra se declara DÍA a DÍA contra 8 h (columna H del
        # registro, con MAX(0,…)), mientras el 01 predica distribución
        # irregular y su 'Cuadrante Mensual' saca la media semanal para
        # compararla con el contrato: diez horas el lunes y seis el martes
        # producen 2 h «extra» que no existen, y los días por debajo de lo
        # contratado no se veían por ninguna parte. Este saldo lleva SIGNO.
        _f(ws, 'H{}'.format(f),
           '=IF($A{f}="","",IF(COUNTIF({e},$A{f}&"")=0,"",'
           'ROUND(SUMIF({e},$A{f}&"",{tr})-SUMIF({e},$A{f}&"",{cn}),2)))'
           .format(f=f, e=emp, tr=tra, cn=con))
        n += 6

    heredadas = 0
    for col in ('B', 'C', 'D', 'F', 'G', 'H'):
        heredadas += _desverdear(ws, '{c}{a}:{c}{b}'.format(c=col, a=r0, b=r1))
    # RT-16 · esas celdas llevaban '0.00' aunque siempre contienen texto.
    for f in range(r0, r1 + 1):
        ws['F{}'.format(f)].number_format = 'General'
    ws['F5'].number_format = 'General'

    fila_tot = _fila_con(ws, 1, 'TOTALES', r1 + 1)
    if fila_tot:
        for col in ('B', 'C', 'D', 'G', 'H'):
            _f(ws, '{}{}'.format(col, fila_tot),
               '=IF(COUNT({c}{a}:{c}{b})=0,"",ROUND(SUM({c}{a}:{c}{b}),2))'
               .format(c=col, a=r0, b=r1))
            ws['{}{}'.format(col, fila_tot)].number_format = (
                motor.FMT_EUR if col == 'G' else motor.FMT_DEC2)
            n += 1
        for col in ('E', 'F'):
            ws['{}{}'.format(col, fila_tot)] = None
    cambios.append('{}:Resumen Mensual!B{}:H{}: {} fórmulas — B agrega por '
                   'SUMIF desde «Registro Horas» (hoy es transcripción '
                   'manual), C descuenta fuerza mayor, compensadas con '
                   'descanso y complementarias (arts. 35.2 y 12.5), F vigila '
                   'el límite anual de B/F3 con semáforo y G usa el recargo '
                   'de D3 como PARÁMETRO. Desaparecen las cabeceras «Coste '
                   '×1.75» y «Coste ×2.0», que presentaban como ley lo que '
                   'fija el convenio — DOM-07/DOM-12/TEC-06/TEC-08/COM-02/'
                   'COM-14/COM-19. {} celdas de las columnas calculadas dejan '
                   'de estar verdes'.format(F02, r0, r1, n, heredadas))
    cambios.append('{}:Resumen Mensual!G{}:G{}: el coste deja de pagar las '
                   'horas COMPENSADAS CON DESCANSO —que por definición del '
                   'art. 35.1 ET se pagan con tiempo, no con dinero: el kit '
                   'las pagaba dos veces— y retribuye las COMPLEMENTARIAS del '
                   'contrato parcial como ordinarias, sin el recargo '
                   '(art. 12.5 ET). Las de fuerza mayor siguen dentro: sólo '
                   'no computan en el tope — RD-02/RT-06/RD-18'
                   .format(F02, r0, r1))
    cambios.append('{}:Resumen Mensual!H{}:H{}: columna nueva «Saldo del mes '
                   '(trabajadas − contratadas)» CON SIGNO. La columna H del '
                   'registro es exceso DIARIO con MAX(0,…): oculta los días '
                   'por debajo de lo contratado, así que con distribución '
                   'irregular no había forma de saber si el exceso estaba '
                   'compensado dentro del periodo — RD-17'
                   .format(F02, r0, r1))
    cambios.append('{}:Resumen Mensual!F5: la cabecera pasa a «Estado frente '
                   'al límite anual» (era el rótulo del parámetro de E3, dos '
                   'filas más arriba, en una columna que devuelve texto) y su '
                   'formato de \'0.00\' a General — RT-16/RC-20'.format(F02))


def _instrucciones_02(wb, cambios):
    lineas = [
        'Cómo usar esta plantilla:',
        '▸ Registra entrada, salida y PAUSA. La pausa es lo que hace '
        'registrable el turno partido: 10:00 → 23:00 con 4 h de pausa son 9 h '
        'trabajadas, no 13.',
        '▸ El turno de NOCHE se registra igual: 23:00 → 07:00 son 8 h. La '
        'fórmula contempla el cruce de medianoche.',
        '▸ El \'Resumen Mensual\' agrega por empleado con SUMIF: no hay que '
        'transcribir nada a mano.',
        motor.AVISO_REGISTRO,
        None,
        'Lo que dice el Estatuto de los Trabajadores:',
        '▸ Art. 35.2: máximo 80 horas extra al año por trabajador. NO computan '
        'en ese tope las compensadas con descanso dentro de los 4 meses '
        'siguientes ni las de fuerza mayor: por eso el registro pide el tipo y '
        'el resumen las descuenta.',
        '▸ Art. 35.1: la hora extra no puede valer MENOS que la ordinaria. La '
        'cuantía exacta la fija tu convenio o tu contrato, y puede compensarse '
        'con descanso en vez de pagarse.',
        '▸ Por eso el recargo es una celda verde (D3) y no un número dentro de '
        'la fórmula. El 1,25 que trae de fábrica es un punto de partida, no '
        'una cifra legal: sustitúyelo por el de tu convenio provincial de '
        'hostelería antes de usar la columna de coste.',
        '▸ Las horas extra son voluntarias salvo pacto en convenio o contrato '
        '(art. 35.4 ET).',
        None,
        'Columnas del registro:',
        '▸ Horas trabajadas = salida − entrada − pausa. Escribe las horas en '
        'formato hh:mm (9:00, no 9): el desplegable rechaza un 9 suelto, que '
        'para Excel son nueve DÍAS y dejaba la jornada en 0,00 sin avisar.',
        '▸ H. Contratadas = jornada CONTRATADA de ese día, precargada a 8. Si '
        'la borras, la hora extra se queda en blanco a propósito: sin saber lo '
        'contratado no hay forma de decir qué sobra.',
        '▸ Tipo: Voluntaria · Obligatoria · Fuerza mayor · Compensada con '
        'descanso · Complementaria (contrato parcial). Las tres últimas no '
        'suman al tope de 80 h.',
        '▸ Las COMPENSADAS CON DESCANSO no se pagan en dinero: se pagan con '
        'tiempo libre (art. 35.1 ET). Por eso la columna de coste las '
        'descuenta — si no, las estarías pagando dos veces. Las de FUERZA '
        'MAYOR sí se pagan: lo que pasa con ellas es que no computan en el '
        'tope anual.',
        '▸ Si el contrato es a tiempo PARCIAL, sus horas de más NO son '
        'extraordinarias —al parcial le están prohibidas (art. 12.4.c ET)—, '
        'son COMPLEMENTARIAS: exigen pacto escrito, tienen su propio tope '
        'sobre la jornada pactada y se retribuyen COMO ORDINARIAS (art. 12.5 '
        'ET). Márcalas con ese tipo y el resumen las deja fuera del tope de '
        '80 h y del recargo.',
        '▸ «Saldo del mes (trabajadas − contratadas)» lleva SIGNO: la columna '
        '«Horas extra» del registro es exceso DIARIO y no ve los días por '
        'debajo de lo contratado. Si distribuyes la jornada de forma '
        'irregular, lo que se paga como extra sale de ESTE saldo al cierre '
        'del periodo de referencia que tengas pactado, no de la suma de los '
        'excesos día a día.',
        '▸ La columna «Aviso» del registro avisa de las cuatro formas de '
        'perder una jornada en silencio: hora escrita sin los dos puntos '
        '(9 en vez de 9:00), entrada igual que salida, pausa mayor que la '
        'jornada y nombre que no está en el «Resumen Mensual» (el SUMIF busca '
        'el texto EXACTO: un espacio de más y esa persona sale a cero).',
        '▸ En el \'Resumen Mensual\', «H. extra acumuladas en el año» es '
        'verde y la llevas tú de mes en mes: suma el acumulado anterior y las '
        'computables de este mes. Es la que vigila el límite.',
        '▸ Antes de usar la tarifa de B3, sácala de la columna «Coste/hora» '
        'del 03: es el coste real por hora, no el bruto del convenio.',
    ]
    _instrucciones(wb['Instrucciones'], lineas)
    cambios.append('{}:Instrucciones: cuerpo reescrito — ×1,75 y ×2,0 dejan '
                   'de presentarse como «Legislación española» (art. 35.1: '
                   'sólo se exige que no valga menos que la ordinaria) y '
                   'entran las dos excepciones del art. 35.2 — DOM-12/COM-19'
                   .format(F02))


# ==========================================================================
# BONUS-01 — caja y temperaturas (DOM-29) + observaciones (TEC-24)
# ==========================================================================
FILAS_BONUS = 19           # 46..64, justo antes de OBSERVACIONES GENERALES

#: RD-22/RD-23 · `(equipo, mínimo, máximo)`.
#:
#: RD-22 · la comparación era UNILATERAL (`B<=C`): una cámara escarchada a
#: −5 °C —producto fresco congelado por accidente, no conformidad APPCC de
#: manual y merma directa— salía «✓ CONFORME» en el documento que firman los
#: dos jefes de turno. Sólo el congelador estaba bien servido por un límite
#: superior a secas. Es la misma corrección que ya se aplicó en el kit de
#: inventario.
#:
#: RD-23 · y faltaba el único control que se hace justo en el relevo con
#: producto en la mano: el mantenimiento EN CALIENTE, que la normativa de
#: comidas preparadas exige a 65 °C o más. El «Expositor / vitrina» se
#: precargaba a 4 °C como si sólo hubiera vitrinas frías.
LIMITES_TEMP = [('Cámara de refrigeración', 0, 4),
                ('Congelador', -30, -18),
                ('Expositor / vitrina FRÍA', 0, 4),
                ('Baño maría / mesa caliente', 65, 90),
                ('Buffet o vitrina CALIENTE', 65, 90)]

#: RD-04 · tolerancia por defecto del descuadre de caja, en euros. Va en celda
#: verde: cada casa tiene la suya.
TOLERANCIA_CAJA = 5.0


def _veredicto_caja(fila, fila_tol):
    """RD-04 — el veredicto del arqueo, con la tolerancia en celda verde.

    `descuadre = efectivo contado − fondo inicial − ventas en efectivo del
    TPV`. Sin la tercera columna no hay arqueo posible: lo que la v2.0
    entregaba era `recaudación − fondo`, la recaudación neta, un número que ya
    se sabe y que un faltante de 80 € no altera.
    """
    return ('=IF($E{f}="","",IF(ABS($E{f})<=$B${t},"✓ CUADRA",'
            'IF($E{f}<0,"⛔ FALTAN "&TEXT(-$E{f},"0.00")&" €",'
            '"⚠ SOBRAN "&TEXT($E{f},"0.00")&" €")))'
            .format(f=fila, t=fila_tol))


def _bonus01(wb, cambios):
    ws = wb['Briefing']
    ws.column_dimensions['A'].width = 20      # 5 era ilegible hasta para
    ws.column_dimensions['B'].width = 34      # «Producto» y «Empleado»
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 26

    fila_obs = _fila_con(ws, 1, 'OBSERVACIONES GENERALES')
    if fila_obs is None:
        return
    if _fila_con(ws, 1, 'CAJA') is None:
        for _ in range(FILAS_BONUS):
            motor.insertar_fila(ws, fila_obs)
        cambios.append('{}:Briefing!{}:{}: {} filas insertadas para los dos '
                       'bloques que faltaban en un traspaso real — DOM-29'
                       .format(FB1, fila_obs, fila_obs + FILAS_BONUS - 1,
                               FILAS_BONUS))
    c0 = _fila_con(ws, 1, 'CAJA')
    if c0 is None:
        c0 = fila_obs

    # ---- CAJA (RD-04) ----------------------------------------------------
    # El bloque se firmaba como ARQUEO y no podía detectar un descuadre, que es
    # la única razón de arquear al cambio de turno: `D = C − B` con B = fondo
    # inicial y C = recaudación es la recaudación NETA, un número que ya se
    # sabe, etiquetado «Diferencia». Sin la lectura Z del TPV, un faltante de
    # 80 € salía exactamente igual que una caja cuadrada.
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    _titulo_bloque(ws, c0, '💶 CAJA: ARQUEO AL CAMBIO DE TURNO', 6)
    _rotulo_tol = ws.cell(row=c0 + 1, column=1,
                          value='Tolerancia de descuadre (€):')
    _rotulo_tol.font = NEGRITA
    _verde_tol = ws.cell(row=c0 + 1, column=2, value=TOLERANCIA_CAJA)
    _verde_tol.number_format = motor.FMT_EUR
    motor.marcar_verde(ws, 'B{}'.format(c0 + 1))
    _limpiar_mis_dv(ws, ('B{}'.format(c0 + 1),))
    _dv(ws, 'B{}'.format(c0 + 1), None, 'Tolerancia no válida',
        'Por debajo de esta cantidad, en más o en menos, el arqueo se da por '
        'cuadrado. Cinco euros es lo habitual en un cambio de turno; súbelo o '
        'bájalo según tu casa.',
        'Escribe un importe entre 0 y 100 €.',
        tipo='decimal', operator='between', formula1='0', formula2='100')
    ws.cell(row=c0 + 1, column=3,
            value='← por debajo de esta cifra el arqueo se da por cuadrado')\
        .font = Font(size=9, italic=True, color='666666')
    _cab(ws, c0 + 2, ['Caja', 'Fondo inicial (€)', 'Efectivo contado (€)',
                      'Ventas en efectivo según TPV — lectura Z (€)',
                      'Descuadre (€)', 'Veredicto'])
    cajas = ('Sala', 'Barra')
    for k, nombre in enumerate(cajas):
        f = c0 + 3 + k
        ws.cell(row=f, column=1, value=nombre)
        _f(ws, 'E{}'.format(f),
           '=IF(OR($B{f}="",$C{f}="",$D{f}=""),"",'
           'ROUND($C{f}-$B{f}-$D{f},2))'.format(f=f))
        _f(ws, 'F{}'.format(f), _veredicto_caja(f, c0 + 1))
        for col in ('B', 'C', 'D', 'E'):
            ws['{}{}'.format(col, f)].number_format = motor.FMT_EUR
    motor.marcar_verde(ws, 'A{}:D{}'.format(c0 + 3, c0 + 2 + len(cajas)))
    ftot = c0 + 3 + len(cajas)
    ws.cell(row=ftot, column=1, value='TOTAL').font = NEGRITA
    for col in ('B', 'C', 'D'):
        _f(ws, '{}{}'.format(col, ftot),
           '=IF(COUNT({c}{a}:{c}{b})=0,"",ROUND(SUM({c}{a}:{c}{b}),2))'
           .format(c=col, a=c0 + 3, b=ftot - 1))
    _f(ws, 'E{}'.format(ftot),
       '=IF(OR($B{f}="",$C{f}="",$D{f}=""),"",'
       'ROUND($C{f}-$B{f}-$D{f},2))'.format(f=ftot))
    _f(ws, 'F{}'.format(ftot), _veredicto_caja(ftot, c0 + 1))
    for col in ('B', 'C', 'D', 'E'):
        cel = ws['{}{}'.format(col, ftot)]
        cel.number_format = motor.FMT_EUR
        cel.font = NEGRITA
    ref_caja = 'F{}:F{}'.format(c0 + 3, ftot)
    motor._limpiar_cf(ws, set([ref_caja]))
    motor.semaforo(ws, ref_caja, motor.VOC_CAJA)
    ffirma = ftot + 1
    ws.merge_cells('A{0}:F{0}'.format(ffirma))
    ws.cell(row=ffirma, column=1,
            value='Arqueo hecho por: _____________________     ·     '
                  'Conforme turno entrante: _____________________')

    # ---- TEMPERATURAS (RD-22/RD-23) --------------------------------------
    t0 = ffirma + 2
    _titulo_bloque(ws, t0, '🌡️ TEMPERATURAS AL CAMBIO DE TURNO (APPCC)', 6)
    _cab(ws, t0 + 1, ['Equipo', 'Temperatura (°C)', 'Mínimo (°C)',
                      'Máximo (°C)', 'Conforme', ''])
    for k, par in enumerate(LIMITES_TEMP):
        f = t0 + 2 + k
        ws.cell(row=f, column=1, value=par[0])
        for col, valor in ((3, par[1]), (4, par[2])):
            cel = ws.cell(row=f, column=col, value=valor)
            cel.number_format = motor.FMT_DEC1
            cel.alignment = CENTRO
        ws.cell(row=f, column=2).number_format = motor.FMT_DEC1
        ws.cell(row=f, column=2).alignment = CENTRO
        _f(ws, 'E{}'.format(f),
           '=IF(OR($B{f}="",$C{f}="",$D{f}=""),"",'
           'IF(AND($B{f}>=$C{f},$B{f}<=$D{f}),"✓ CONFORME",'
           '"⛔ FUERA DE RANGO"))'.format(f=f))
    ult_temp = t0 + 1 + len(LIMITES_TEMP)
    motor.marcar_verde(ws, 'B{}:D{}'.format(t0 + 2, ult_temp))
    ref_cf = 'E{}:E{}'.format(t0 + 2, ult_temp)
    motor._limpiar_cf(ws, set([ref_cf]))
    motor.semaforo(ws, ref_cf, motor.VOC_CONFORME)
    fnota = ult_temp + 1
    ws.merge_cells('A{0}:F{0}'.format(fnota))
    cel = ws.cell(row=fnota, column=1,
                  value='Quien entrega toma la temperatura DELANTE de quien '
                        'recibe: es el punto en que el APPCC cambia de '
                        'responsable. El rango tiene MÍNIMO y máximo: una '
                        'cámara a −5 °C también es una no conformidad (te ha '
                        'congelado el producto fresco), y el mantenimiento en '
                        'caliente sólo es conforme A PARTIR de 65 °C. Fuera '
                        'de rango, anótalo arriba como incidencia y avisa '
                        'antes de irte.')
    cel.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[fnota].height = 42

    # ---- observaciones: wrap_text y alto (TEC-24) ------------------------
    obs = _fila_con(ws, 1, 'OBSERVACIONES GENERALES')
    for f in range(obs, obs + 4):
        for c in range(1, 5):
            ws.cell(row=f, column=c).alignment = Alignment(wrap_text=True,
                                                            vertical='top')
        if f > obs:
            ws.row_dimensions[f].height = 30
    motor.marcar_verde(ws, 'A{}:D{}'.format(obs + 1, obs + 3))

    _instrucciones(wb['Instrucciones'], [
        'Cómo usar esta plantilla:',
        '▸ El turno saliente completa este briefing ANTES de irse.',
        '▸ El turno entrante lo revisa y firma al llegar.',
        '▸ Imprime una copia diaria o úsalo en la tablet.',
        '▸ Archiva los briefings: son la trazabilidad del traspaso, y las '
        'temperaturas de abajo son además registro de autocontrol APPCC.',
        None,
        'Secciones incluidas:',
        '▸ Reservas pendientes y VIPs del siguiente turno.',
        '▸ Incidencias: averías, falta de stock, quejas. La gravedad va con '
        'desplegable (Alta · Media · Baja): sin escala fija cada encargado '
        'escribe una cosa y el archivo deja de ser comparable.',
        '▸ Tareas pendientes que no se completaron, con prioridad (Urgente · '
        'Alta · Normal).',
        '▸ Stock bajo que necesita pedido urgente.',
        '▸ Personal: ausencias y cambios de último momento.',
        '▸ CAJA — arqueo de verdad: fondo inicial, EFECTIVO CONTADO y VENTAS '
        'EN EFECTIVO según la lectura Z del TPV. El descuadre es contado − '
        'fondo − ventas, y el veredicto lo compara con la tolerancia que '
        'fijes arriba (5 € de fábrica). Sin la columna del TPV esto no sería '
        'un arqueo: sería la recaudación neta, un número que ya sabes y que '
        'un faltante de 80 € no cambia.',
        '▸ Un FALTANTE sale en rojo y un SOBRANTE en ámbar: los dos son '
        'descuadres, pero un sobrante suele ser un cobro sin registrar y no '
        'un agujero de caja. Los dos se firman abajo.',
        '▸ TEMPERATURAS al cambio de turno: cámara, congelador, vitrina FRÍA, '
        'baño maría / mesa caliente y buffet o vitrina CALIENTE. Es el punto '
        'exacto en que el APPCC cambia de responsable. Cada equipo tiene '
        'MÍNIMO y máximo, no sólo un techo: una cámara a −5 °C te ha '
        'congelado el producto fresco y es no conformidad igual que una a '
        '10 °C, y el mantenimiento en caliente sólo es conforme A PARTIR de '
        '65 °C. La columna «Conforme» se pinta sola.',
        '▸ Observaciones generales: las tres líneas admiten texto largo (van '
        'con ajuste de línea y alto de fila, para que no se corten al '
        'imprimir).',
    ])
    cambios.append('{}:Briefing!A{}:F{}: bloque de ARQUEO DE CAJA (fondo · '
                   'efectivo contado · ventas en efectivo del TPV · '
                   'DESCUADRE · veredicto contra una tolerancia en celda '
                   'verde · firma) y bloque de TEMPERATURAS con MÍNIMO y '
                   'máximo para cinco equipos —cámara 0/4, congelador '
                   '−30/−18, vitrina fría 0/4 y los dos de mantenimiento en '
                   'CALIENTE a 65/90— con semáforo CONFORME / FUERA DE '
                   'RANGO; observaciones con wrap_text y alto de fila. Antes '
                   'la «Diferencia» era recaudación − fondo (no detectaba '
                   'ningún descuadre) y la conformidad sólo miraba el límite '
                   'SUPERIOR (una cámara a −5 °C salía «✓ CONFORME») — '
                   'DOM-29/TEC-24/RD-04/RD-22/RD-23'
                   .format(FB1, c0, obs + 3))


# ==========================================================================
# Interfaz del pipeline
# ==========================================================================
def pre(wb, fname, cambios):
    if fname == F02:
        _pre_02(wb, cambios)


def post(wb, fname, cambios, registro_grupo):
    if fname == F01:
        _hoja_turnos(wb, cambios)
        r0, _r1 = _cuadrante_semanal(wb['Cuadrante Semanal'], cambios)
        _cuadrante_mensual(wb, cambios, r0)
        _instrucciones_01(wb, cambios)
    elif fname == F02:
        reg0, reg1 = _registro_horas(wb['Registro Horas'], cambios)
        _resumen_mensual(wb['Resumen Mensual'], cambios, reg0, reg1)
        _instrucciones_02(wb, cambios)
    elif fname == FB1:
        _bonus01(wb, cambios)


# ==========================================================================
# Demostraciones con pycel (SPEC §5)
# ==========================================================================
def _pycel_helpers():
    """Reutiliza los helpers de `main.py` cuando el módulo corre dentro del
    orquestador (que es siempre); el respaldo local sólo existe para poder
    importar este módulo suelto."""
    m = sys.modules.get('__main__')
    if m is not None and all(hasattr(m, k) for k in ('_ev', '_set', '_pycel',
                                                     'hora')):
        return m._pycel, m._ev, m._set, m.hora
    import contextlib

    def _pycel(path):
        from pycel import ExcelCompiler
        return ExcelCompiler(filename=path)

    def _ev(xl, ref):
        with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
            try:
                return xl.evaluate(ref)
            except Exception:                              # noqa: BLE001
                return 'ERR'

    def _set(xl, ref, valor):
        with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
            try:
                xl.evaluate(ref)
                xl.set_value(ref, valor)
                return True
            except Exception:                              # noqa: BLE001
                return False

    def hora(hh, mm=0):
        return (hh * 60.0 + mm) / 1440.0
    return _pycel, _ev, _set, hora


def demos(carpeta, origen):
    """Cada bloque CAMBIA una entrada y comprueba la DIRECCIÓN del resultado.
    Un valor correcto con la hoja vacía no demuestra nada.

    Cada escenario usa un compilador de pycel **limpio** (`ExcelCompiler`
    cuesta 0,16 s sobre el 01, medido). No es celo: pycel **no propaga la
    invalidación de `set_value` a través de un nodo RANGO**, así que reutilizar
    un compilador para escenarios encadenados devuelve valores de la ronda
    anterior — en la primera ejecución de esta batería, `Resumen Mensual!F6`
    llegó a informar «⛔ EXCEDE (25 h)» con la celda de acumulado VACÍA. Un
    gate que arrastra estado no es un gate.
    """
    _pyc, _ev, _set, _hora = _pycel_helpers()

    def esc(path, sets, lecturas):
        """Compilador limpio → aplica las entradas → lee las salidas."""
        xl = _pyc(path)
        for ref, val in sets:
            _set(xl, ref, val)
        return dict((k, _ev(xl, r)) for k, r in lecturas)

    def sin_cache(path):
        """Copia del libro SIN valores cacheados, para poder medir cadenas
        que pasan por una celda con cache.

        pycel devuelve el valor CACHEADO de una celda de fórmula si el libro lo
        trae, y `main.py` corre `inject_cache` ANTES de las demostraciones. Con
        `Turnos!E5` cacheada en 8, mover `Turnos!D5` no cambiaba nada aguas
        abajo por más que la fórmula fuera correcta: el gate no medía el
        arreglo de RT-02, medía el cache. Volver a guardar el libro con
        openpyxl BORRA todos los `<v>` —es exactamente el motivo por el que
        `inject_cache` va siempre el último— y deja la cadena viva.

        La copia va a un directorio HERMANO del de trabajo: nada de dejar un
        décimo .xlsx dentro de la carpeta que audita el censo.
        """
        destino = os.path.join(os.environ.get('CLAUDE_SCRATCHPAD', os.path.dirname(carpeta)), 'demos-sin-cache')
        if not os.path.isdir(destino):
            os.makedirs(destino)
        fuera_path = os.path.join(destino, os.path.basename(path))
        openpyxl.load_workbook(path).save(fuera_path)
        return fuera_path

    fuera = {}
    p01 = os.path.join(carpeta, F01)
    p02 = os.path.join(carpeta, F02)
    pb1 = os.path.join(carpeta, FB1)
    cs = "'Cuadrante Semanal'!"

    # ---- 01 · las 4 alertas ---------------------------------------------
    if os.path.isfile(p01):
        def plan(codigos, contratadas=40, nombre='Demo'):
            sets = [(cs + '{}6'.format(get_column_letter(2 + j)), c)
                    for j, c in enumerate(codigos)]
            return sets + [(cs + 'J6', contratadas), (cs + 'A6', nombre)]

        L = ('I', 'K', 'L', 'M', 'N')
        lect = [(c, cs + '{}6'.format(c)) for c in L]
        vacia = esc(p01, [], lect)
        tm = esc(p01, plan(['T', 'M', 'L', 'L', 'L', 'L', 'L']), lect)
        mm = esc(p01, plan(['M', 'M', 'L', 'L', 'L', 'L', 'L']), lect)
        nt = esc(p01, plan(['N', 'T', 'L', 'L', 'L', 'L', 'L']), lect)
        siete = esc(p01, plan(['M'] * 7), lect)
        seis = esc(p01, plan(['M'] * 6 + ['L']), lect)
        cinco = esc(p01, plan(['M'] * 5 + ['L', 'L']), lect)
        parcial = esc(p01, plan(['M'] * 4 + ['L', 'L', 'L'], contratadas=20),
                      lect)
        doble = esc(p01, plan(['D', 'L', 'L', 'L', 'L', 'L', 'L']), lect)
        partido = esc(p01, plan(['P', 'L', 'L', 'L', 'L', 'L', 'L']), lect)

        pruebas = [
            {'alerta': 'K · descanso entre jornadas (art. 34.3 ET, 12 h)',
             'ref': '{}:Cuadrante Semanal:K6'.format(F01),
             'hoja recién descargada': vacia['K'],
             'T→M · 8 h de descanso': tm['K'],
             'M→M · 16 h de descanso': mm['K'],
             'N→T · el turno cruza la medianoche, 8 h': nt['K'],
             'ok': (vacia['K'] in ('', None) and bool(tm['K'])
                    and mm['K'] in ('', None) and bool(nt['K']))},
            {'alerta': 'L · descanso semanal (art. 37.1 ET)',
             'ref': '{}:Cuadrante Semanal:L6'.format(F01),
             'hoja recién descargada': vacia['L'],
             '7 días con turno': siete['L'], '1 día libre': seis['L'],
             '2 días libres': cinco['L'],
             'ok': (vacia['L'] in ('', None) and '0 días' in str(siete['L'])
                    and '1 día' in str(seis['L'])
                    and cinco['L'] in ('', None))},
            {'alerta': 'M · jornada semanal contra lo CONTRATADO (DOM-25)',
             'ref': '{}:Cuadrante Semanal:M6'.format(F01),
             'hoja recién descargada': vacia['M'],
             '40 h planificadas / 40 contratadas': cinco['M'],
             'horas de esa semana (I6)': cinco['I'],
             '32 h planificadas / 20 contratadas': parcial['M'],
             'ok': (vacia['M'] in ('', None) and cinco['M'] in ('', None)
                    and cinco['I'] == 40 and '+12' in str(parcial['M'])
                    and parcial['I'] == 32)},
            {'alerta': 'N · jornada diaria (art. 34.3 ET, 9 h)',
             'ref': '{}:Cuadrante Semanal:N6'.format(F01),
             'hoja recién descargada': vacia['N'],
             'turno DOBLE de 16 h': doble['N'],
             'turno PARTIDO de 9 h': partido['N'],
             'ok': (vacia['N'] in ('', None) and bool(doble['N'])
                    and partido['N'] in ('', None))},
        ]
        fuera['grupo_a_4_alertas_01'] = {
            'ref': '{}:Cuadrante Semanal:K6:N6'.format(F01),
            'pruebas': pruebas, 'ok': all(p['ok'] for p in pruebas),
            'nota': 'las cuatro alertas que la landing vende desde la v1.0 y '
                    'que el fichero no tenía: la única era «>40 h» con el 40 '
                    'escrito dentro de la fórmula (COM-01/DOM-25)'}

        # el umbral es un PARÁMETRO, no un literal (§1.4)
        base = plan(['M', 'M', 'L', 'L', 'L', 'L', 'L'])
        con12 = esc(p01, base, [('K', cs + 'K6')])
        con18 = esc(p01, base + [('Turnos!B3', 18)], [('K', cs + 'K6')])
        fuera['grupo_a_parametro_descanso_01'] = {
            'ref': '{}:Turnos:B3 → Cuadrante Semanal:K6'.format(F01),
            'M→M con el mínimo legal de 12 h': con12['K'],
            'M→M con un convenio que exija 18 h': con18['K'],
            'ok': con12['K'] in ('', None) and bool(con18['K']),
            'nota': 'M→M deja 16 h de descanso: cumple las 12 del art. 34.3 y '
                    'no cumpliría un convenio de 18. El umbral vive en celda, '
                    'no dentro de la fórmula (§1.4)'}

        # RT-02 · la tabla `Turnos` manda sobre las horas Y las horas salen del
        # HORARIO. La versión anterior tenía la columna E como constante verde:
        # cambiar `Turnos!D5` (la hora de fin de la mañana) dejaba `I6` en 48,0
        # y `N6` callada aunque la jornada real pasara de 10 h. Ahora se mide
        # moviendo C y D, que es lo que prometen las Instrucciones.
        uno = plan(['M', 'L', 'L', 'L', 'L', 'L', 'L'])
        p01v = sin_cache(p01)          # `Turnos!E` viene cacheada: ver arriba
        lect_in = [('E', 'Turnos!E5'), ('I', cs + 'I6'), ('N', cs + 'N6')]
        h8 = esc(p01v, uno, lect_in)
        fin17 = esc(p01v, uno + [('Turnos!D5', 17)], lect_in)  # 7→17 = 10 h
        fin13 = esc(p01v, uno + [('Turnos!D5', 13)], lect_in)  # 7→13 = 6 h
        ini8 = esc(p01v, uno + [('Turnos!C5', 8)], lect_in)    # 8→15 = 7 h
        # el partido: 10→23 con 4 h de pausa siguen siendo 9 h efectivas
        pausa = esc(p01v, plan(['P', 'L', 'L', 'L', 'L', 'L', 'L']),
                    [('E', 'Turnos!E8'), ('I', cs + 'I6')])
        sin_pausa = esc(p01v, plan(['P', 'L', 'L', 'L', 'L', 'L', 'L'])
                        + [('Turnos!F8', 0)],
                        [('E', 'Turnos!E8'), ('I', cs + 'I6')])
        fuera['grupo_a_turnos_manda_01'] = {
            'ref': '{}:Turnos:C5/D5/F8 → Turnos:E → Cuadrante Semanal:I6/N6'
                   .format(F01),
            'M de fábrica (7→15, sin pausa)': {'E5': h8['E'], 'I6': h8['I'],
                                               'N6': h8['N']},
            'M con la hora de FIN a las 17 (10 h > 9 del máximo)':
                {'E5': fin17['E'], 'I6': fin17['I'], 'N6': fin17['N']},
            'M con la hora de FIN a las 13': {'E5': fin13['E'],
                                              'I6': fin13['I']},
            'M empezando a las 8': {'E5': ini8['E'], 'I6': ini8['I']},
            'P de fábrica (10→23 con 4 h de pausa)': {'E8': pausa['E'],
                                                      'I6': pausa['I']},
            'P sin pausa (13 h entre extremos)': {'E8': sin_pausa['E'],
                                                  'I6': sin_pausa['I']},
            'ok': (h8['E'] == 8 and h8['I'] == 8 and h8['N'] in ('', None)
                   and fin17['E'] == 10 and fin17['I'] == 10
                   and bool(fin17['N'])
                   and fin13['E'] == 6 and fin13['I'] == 6
                   and ini8['E'] == 7 and ini8['I'] == 7
                   and pausa['E'] == 9 and pausa['I'] == 9
                   and sin_pausa['E'] == 13 and sin_pausa['I'] == 13),
            'nota': 'las horas salen de la tabla —no de los siete IF anidados '
                    'de 700 caracteres de la v1.1 (DOM-05)— y la columna '
                    '«Horas» sale a su vez del horario menos la pausa: mover '
                    'la hora de fin recalcula el total del cuadrante Y '
                    'despierta la alerta de jornada diaria (RT-02)'}

        # RD-16 · menores: jornada de 8 h y prohibición de nocturno y doble
        men_no = esc(p01, plan(['M', 'M', 'M', 'M', 'M', 'L', 'L']),
                     [('N', cs + 'N6'), ('P', cs + 'P6')])
        men_si = esc(p01, plan(['M', 'M', 'M', 'M', 'M', 'L', 'L'])
                     + [(cs + 'O6', 'S')],
                     [('N', cs + 'N6'), ('P', cs + 'P6')])
        men_9h = esc(p01, plan(['P', 'L', 'L', 'L', 'L', 'L', 'L'])
                     + [(cs + 'O6', 'S')],
                     [('N', cs + 'N6'), ('P', cs + 'P6')])
        men_noche = esc(p01, plan(['N', 'L', 'L', 'L', 'L', 'L', 'L'])
                        + [(cs + 'O6', 'S')],
                        [('N', cs + 'N6'), ('P', cs + 'P6')])
        men_doble = esc(p01, plan(['D', 'L', 'L', 'L', 'L', 'L', 'L'])
                        + [(cs + 'O6', 'S')],
                        [('N', cs + 'N6'), ('P', cs + 'P6')])
        fuera['grupo_a_menores_01'] = {
            'ref': '{}:Cuadrante Semanal:O6 → N6/P6'.format(F01),
            'mayor de edad, 5 mañanas de 8 h': {'N': men_no['N'],
                                                'P': men_no['P']},
            'MENOR, 5 mañanas de 8 h': {'N': men_si['N'], 'P': men_si['P']},
            'MENOR con un PARTIDO de 9 h (legal para un adulto)':
                {'N': men_9h['N'], 'P': men_9h['P']},
            'MENOR con turno de NOCHE': {'N': men_noche['N'],
                                         'P': men_noche['P']},
            'MENOR con turno DOBLE': {'N': men_doble['N'],
                                      'P': men_doble['P']},
            'ok': (men_no['N'] in ('', None) and men_no['P'] in ('', None)
                   and men_si['N'] in ('', None)
                   and 'sin horas extra' in str(men_si['P'])
                   and 'MENOR' in str(men_9h['N'])
                   and 'prohibidos' in str(men_noche['P'])
                   and 'prohibidos' in str(men_doble['P'])),
            'nota': 'el único aviso de minoría de edad del kit vivía en el 07 '
                    'y no llegaba al 01, que es donde se comete la infracción: '
                    'a un menor se le podía poner N o D sin que ninguna de las '
                    'cuatro alertas dijera nada (art. 6 ET) — RD-16'}

        # la hoja mensual bebe del semanal y calcula (DOM-17/COM-12)
        men = esc(p01, [(cs + 'A7', 'Marta Ibáñez'), (cs + 'J7', 40),
                        ("'Cuadrante Mensual'!B6", 'M'),
                        ("'Cuadrante Mensual'!C6", 'D'),
                        ("'Cuadrante Mensual'!D6", 'D')],
                  [('nombre', "'Cuadrante Mensual'!A6"),
                   ('horas', "'Cuadrante Mensual'!I6"),
                   ('alerta', "'Cuadrante Mensual'!J6"),
                   ('sem1_del_total', "'Cuadrante Mensual'!B171"),
                   ('total_mes', "'Cuadrante Mensual'!G171")])
        fuera['grupo_a_mensual_vivo_01'] = {
            'ref': '{}:Cuadrante Mensual:A6/I6/J6 y G171'.format(F01),
            'nombre heredado de Cuadrante Semanal!A7': men['nombre'],
            'M + D + D = 8 + 16 + 16': men['horas'],
            'alerta contra las 40 h contratadas': men['alerta'],
            'esa semana en el total del mes': men['sem1_del_total'],
            'total del mes': men['total_mes'],
            'ok': (men['nombre'] == 'Marta Ibáñez' and men['horas'] == 40
                   and men['alerta'] in ('', None)
                   and men['sem1_del_total'] == 40
                   and men['total_mes'] == 40),
            'nota': 'la hoja mensual tenía 30 rótulos, CERO fórmulas, CERO '
                    'validación y 28 días (DOM-17/TEC-23/COM-12)'}

    # ---- 02 · medianoche, pausa, guarda y límite -------------------------
    if os.path.isfile(p02):
        rh = "'Registro Horas'!"
        casos = []
        for etiqueta, ent, sal, pausa, contratadas, esperado in (
                ('turno de noche 23:00 → 07:00', (23, 0), (7, 0), 0, 8, 8.0),
                ('cierre 19:00 → 01:30', (19, 0), (1, 30), 0, 8, 6.5),
                ('partido 10:00 → 23:00 con 4 h de pausa', (10, 0), (23, 0), 4,
                 8, 9.0)):
            r = esc(p02, [(rh + 'C5', _hora(*ent)), (rh + 'D5', _hora(*sal)),
                          (rh + 'E5', pausa), (rh + 'G5', contratadas)],
                    [('horas', rh + 'F5'), ('extra', rh + 'H5')])
            casos.append({'caso': etiqueta, 'esperado_horas': esperado,
                          'horas': r['horas'], 'extra': r['extra'],
                          'ok': isinstance(r['horas'], (int, float))
                          and abs(r['horas'] - esperado) < 0.005})
        jornada = [(rh + 'C5', _hora(9)), (rh + 'D5', _hora(17)),
                   (rh + 'E5', 0)]
        sin_g = esc(p02, jornada + [(rh + 'G5', '')],
                    [('extra', rh + 'H5'), ('horas', rh + 'F5')])
        con_g = esc(p02, jornada + [(rh + 'G5', 6)], [('extra', rh + 'H5')])
        igual = esc(p02, jornada + [(rh + 'G5', 8)], [('extra', rh + 'H5')])
        casos.append({'caso': 'H. contratadas VACÍA con 8 h trabajadas',
                      'horas': sin_g['horas'], 'extra': sin_g['extra'],
                      'ok': sin_g['extra'] in ('', None)
                      and sin_g['horas'] == 8})
        casos.append({'caso': '8 h trabajadas contra 6 contratadas',
                      'extra': con_g['extra'],
                      'ok': isinstance(con_g['extra'], (int, float))
                      and abs(con_g['extra'] - 2) < 0.005})
        casos.append({'caso': '8 h trabajadas contra 8 contratadas',
                      'extra': igual['extra'],
                      'ok': isinstance(igual['extra'], (int, float))
                      and abs(igual['extra']) < 0.005})
        fuera['grupo_a_horas_y_guarda_02'] = {
            'ref': '{}:Registro Horas:F5/H5'.format(F02),
            'pruebas': casos, 'ok': all(c['ok'] for c in casos),
            'nota': 'la v1.1 hacía (D5−C5)*24 y el turno de noche fichaba '
                    '−16 h; y con la columna de contratadas vacía declaraba '
                    'extra la jornada ENTERA — 8 h × 12 € × 1,25 = 120 € por '
                    'día y empleado que el cliente no debe '
                    '(DOM-03/DOM-11/DOM-21/TEC-01/TEC-09/COM-07)'}

        # el resumen AGREGA, descuenta el art. 35.2 y usa el recargo en celda
        registros = [
            (rh + 'A5', 'Ana Prieto'), (rh + 'C5', _hora(9)),
            (rh + 'D5', _hora(17)), (rh + 'E5', 0), (rh + 'G5', 6),
            (rh + 'I5', 'Voluntaria'),
            (rh + 'A6', 'Ana Prieto'), (rh + 'C6', _hora(9)),
            (rh + 'D6', _hora(20)), (rh + 'E6', 0), (rh + 'G6', 8),
            (rh + 'I6', 'Fuerza mayor'),
            (rh + 'A7', 'Luis Cabo'), (rh + 'C7', _hora(9)),
            (rh + 'D7', _hora(19)), (rh + 'E7', 0), (rh + 'G7', 8),
            (rh + 'I7', 'Voluntaria'),
            ("'Resumen Mensual'!A6", 'Ana Prieto'),
            ("'Resumen Mensual'!A7", 'Luis Cabo')]
        lect = [('B', "'Resumen Mensual'!B6"), ('C', "'Resumen Mensual'!C6"),
                ('D', "'Resumen Mensual'!D6"), ('G', "'Resumen Mensual'!G6"),
                ('otro', "'Resumen Mensual'!B7")]
        r125 = esc(p02, registros, lect)
        r150 = esc(p02, registros + [("'Resumen Mensual'!D3", 1.5)], lect)
        r20 = esc(p02, registros + [("'Resumen Mensual'!B3", 20.0)], lect)
        fuera['grupo_a_resumen_agrega_02'] = {
            'ref': '{}:Resumen Mensual:B6/C6/D6/G6'.format(F02),
            'extra de Ana en el mes (2 h + 3 h)': r125['B'],
            'de las que NO computan (fuerza mayor, art. 35.2)': r125['C'],
            'computables en el tope de 80 h': r125['D'],
            'no mezcla empleados: extra de Luis (2 h)': r125['otro'],
            'coste con recargo 1,25 y tarifa 12 €': r125['G'],
            'coste subiendo el recargo a 1,50': r150['G'],
            'coste subiendo la tarifa a 20 €': r20['G'],
            'ok': (r125['B'] == 5 and r125['C'] == 3 and r125['D'] == 2
                   and r125['otro'] == 2
                   and abs(r125['G'] - 75) < 0.01
                   and abs(r150['G'] - 90) < 0.01
                   and abs(r20['G'] - 125) < 0.01),
            'nota': 'la columna B era transcripción MANUAL (COM-14) y el '
                    '×1,75 / ×2,0 estaba escrito dentro de 30 fórmulas '
                    '(TEC-06); ahora los dos son celda verde'}

        # RD-02/RT-06/RD-18 · qué se PAGA y qué computa en el tope
        rm = "'Resumen Mensual'!"

        def coste(tipo, contratadas=8, salida=21):
            sets = [(rh + 'A5', 'Sara Gil'), (rh + 'C5', _hora(9)),
                    (rh + 'D5', _hora(salida)), (rh + 'E5', 0),
                    (rh + 'G5', contratadas), (rh + 'I5', tipo),
                    (rm + 'A6', 'Sara Gil')]
            return esc(p02, sets, [('B', rm + 'B6'), ('C', rm + 'C6'),
                                   ('D', rm + 'D6'), ('G', rm + 'G6')])

        vol = coste('Voluntaria')
        fmay = coste('Fuerza mayor')
        comp = coste(TIPO_COMPENSADA)
        cplm = coste(TIPO_COMPLEMENTARIA)
        fuera['grupo_a_coste_por_tipo_02'] = {
            'ref': '{}:Resumen Mensual:C6/D6/G6 (4 h extra a 12,00 € × 1,25)'
                   .format(F02),
            'Voluntaria': {'total': vol['B'], 'no_computables': vol['C'],
                           'computables': vol['D'], 'coste': vol['G']},
            'Fuerza mayor (se paga; no computa en el tope)':
                {'no_computables': fmay['C'], 'computables': fmay['D'],
                 'coste': fmay['G']},
            'Compensada con descanso (NO se paga en dinero)':
                {'no_computables': comp['C'], 'computables': comp['D'],
                 'coste': comp['G']},
            'Complementaria (contrato parcial: se paga SIN recargo)':
                {'no_computables': cplm['C'], 'computables': cplm['D'],
                 'coste': cplm['G']},
            'ok': (vol['B'] == 4 and vol['C'] == 0 and vol['D'] == 4
                   and abs(vol['G'] - 60) < 0.01
                   and fmay['C'] == 4 and fmay['D'] == 0
                   and abs(fmay['G'] - 60) < 0.01
                   and comp['C'] == 4 and comp['D'] == 0
                   and abs(comp['G']) < 0.01
                   and cplm['C'] == 4 and cplm['D'] == 0
                   and abs(cplm['G'] - 48) < 0.01),
            'nota': 'antes G se calculaba sobre B —TODAS las horas extra— y '
                    'facturaba a 1,25× las COMPENSADAS CON DESCANSO, que por '
                    'definición del art. 35.1 ET se pagan con tiempo y no con '
                    'dinero: 4 h imputaban 60 € que nadie desembolsa y encima '
                    'ya se había dado el descanso. Las de fuerza mayor SÍ se '
                    'pagan (60 €) y sólo salen del tope; las complementarias '
                    'del contrato parcial se pagan como ORDINARIAS, sin el '
                    '1,25 (48 €, art. 12.5 ET) — RD-02/RT-06/RD-18'}

        # RD-17 · saldo NETO del mes, con signo
        def saldo(dias):
            sets = [(rm + 'A6', 'Iker Sanz')]
            for i, (sal, contr) in enumerate(dias):
                f = 5 + i
                sets += [(rh + 'A{}'.format(f), 'Iker Sanz'),
                         (rh + 'C{}'.format(f), _hora(9)),
                         (rh + 'D{}'.format(f), _hora(sal)),
                         (rh + 'E{}'.format(f), 0),
                         (rh + 'G{}'.format(f), contr)]
            return esc(p02, sets, [('extra', rm + 'B6'),
                                   ('saldo', rm + 'H6')])

        irregular = saldo([(19, 8), (15, 8)])      # 10 h + 6 h contra 8 y 8
        corto = saldo([(15, 8), (15, 8)])          # 6 h + 6 h contra 8 y 8
        fuera['grupo_a_saldo_neto_02'] = {
            'ref': '{}:Resumen Mensual:H6'.format(F02),
            'lunes 10 h y martes 6 h, 8 contratadas cada día':
                {'horas extra declaradas (columna B)': irregular['extra'],
                 'saldo NETO del mes (columna H)': irregular['saldo']},
            'dos días de 6 h contra 8 contratadas':
                {'horas extra declaradas': corto['extra'],
                 'saldo NETO': corto['saldo']},
            'ok': (irregular['extra'] == 2 and abs(irregular['saldo']) < 0.005
                   and corto['extra'] == 0 and abs(corto['saldo'] + 4) < 0.005),
            'nota': 'la columna «Horas extra» es exceso DIARIO con MAX(0,…): '
                    'diez horas el lunes y seis el martes producen 2 h extra '
                    'que no existen —la jornada se compensa dentro de la '
                    'semana— y los días por debajo de lo contratado no se '
                    'veían en ningún sitio. El saldo neto sale 0 en el primer '
                    'caso y −4 en el segundo (RD-17)'}

        # RD-19/RT-05/RT-09/RT-15 · las cuatro formas de perder una jornada
        def aviso(sets):
            return esc(p02, sets, [('F', rh + 'F5'),
                                   ('J', rh + COL_AVISO_REG + '5'),
                                   ('H', rh + 'H5')])

        num = aviso([(rh + 'C5', 9), (rh + 'D5', 17), (rh + 'G5', 8)])
        ig = aviso([(rh + 'C5', _hora(9)), (rh + 'D5', _hora(9)),
                    (rh + 'G5', 8)])
        pau = aviso([(rh + 'C5', _hora(9)), (rh + 'D5', _hora(13)),
                     (rh + 'E5', 6), (rh + 'G5', 8)])
        nom = aviso([(rh + 'A5', 'Ana Pérez '), (rh + 'C5', _hora(9)),
                     (rh + 'D5', _hora(21)), (rh + 'E5', 0), (rh + 'G5', 8),
                     ("'Resumen Mensual'!A6", 'Ana Pérez')])
        bien = aviso([(rh + 'A5', 'Ana Pérez'), (rh + 'C5', _hora(9)),
                      (rh + 'D5', _hora(21)), (rh + 'E5', 0), (rh + 'G5', 8),
                      ("'Resumen Mensual'!A6", 'Ana Pérez')])
        fuera['grupo_a_avisos_registro_02'] = {
            'ref': '{}:Registro Horas:F5/{}5'.format(F02, COL_AVISO_REG),
            'hora tecleada como número (9 y 17)': {'F': num['F'],
                                                   'aviso': num['J']},
            'entrada igual que salida (9:00 y 9:00)': {'F': ig['F'],
                                                       'aviso': ig['J']},
            'pausa de 6 h en una jornada de 4': {'F': pau['F'],
                                                 'aviso': pau['J']},
            'nombre con un espacio final que el SUMIF no encuentra':
                {'F': nom['F'], 'extra': nom['H'], 'aviso': nom['J']},
            'la misma fila con el nombre correcto': {'F': bien['F'],
                                                     'aviso': bien['J']},
            'ok': (num['F'] in ('', None) and 'dos puntos' in str(num['J'])
                   and ig['F'] in ('', None) and 'iguales' in str(ig['J'])
                   and pau['F'] in ('', None) and 'pausa' in str(pau['J'])
                   and nom['F'] == 12 and 'no está' in str(nom['J'])
                   and bien['F'] == 12 and bien['J'] in ('', None)),
            'nota': 'antes las tres primeras devolvían 0,00 h o −2,00 h en '
                    'silencio y la cuarta hacía desaparecer las horas del '
                    'cómputo y del coste sin un solo aviso, en el documento '
                    'que hay que conservar cuatro años (art. 34.9 ET) — '
                    'RD-19/RT-05/RT-09/RT-15'}

        # el contador del límite anual, con el tope en celda
        def lim(acum, tope=None):
            sets = [("'Resumen Mensual'!E6", acum)]
            if tope is not None:
                sets.append(("'Resumen Mensual'!F3", tope))
            return esc(p02, sets, [('F', "'Resumen Mensual'!F6")])['F']
        v_vacio, v40, v70, v95, v95_120 = (lim(''), lim(40), lim(70), lim(95),
                                           lim(95, 120))
        fuera['grupo_a_limite_80h_02'] = {
            'ref': '{}:Resumen Mensual:F6 (tope en F3)'.format(F02),
            'recién descargado, sin acumulado': v_vacio,
            'acumuladas 40 h': v40, 'acumuladas 70 h': v70,
            'acumuladas 95 h': v95,
            'acumuladas 95 h con el tope subido a 120': v95_120,
            'ok': (v_vacio in ('', None) and 'dentro' in str(v40)
                   and 'cerca' in str(v70) and 'EXCEDE' in str(v95)
                   and 'dentro' in str(v95_120)),
            'nota': 'art. 35.2 ET: 80 h/año. El tope es celda verde porque se '
                    'prorratea en los contratos parciales y en los de '
                    'duración inferior al año'}

    # ---- BONUS-01 · caja y temperaturas ---------------------------------
    if os.path.isfile(pb1):
        wsb = openpyxl.load_workbook(pb1)['Briefing']
        fc = _fila_con(wsb, 1, 'CAJA')
        ft = _fila_con(wsb, 1, 'TEMPERATURAS')
        if fc and ft:
            # RD-04 · el bloque de caja tiene ahora cabecera en `fc+2` y datos
            # a partir de `fc+3`; el de temperaturas, cabecera en `ft+1`.
            b, t = fc + 3, ft + 2
            tot = b + 2
            bri = "'Briefing'!"

            def arqueo(fondo, contado, tpv, tol=None):
                sets = [(bri + 'B{}'.format(b), fondo),
                        (bri + 'C{}'.format(b), contado),
                        (bri + 'D{}'.format(b), tpv)]
                if tol is not None:
                    sets.append((bri + 'B{}'.format(fc + 1), tol))
                return esc(pb1, sets, [('desc', bri + 'E{}'.format(b)),
                                       ('ver', bri + 'F{}'.format(b))])

            cuadra = arqueo(200, 1650.50, 1450.50)
            falta = arqueo(200, 1570.50, 1450.50)     # faltan 80 €
            sobra = arqueo(200, 1660.50, 1450.50)     # sobran 10 €
            dentro = arqueo(200, 1653.50, 1450.50)    # +3 €, dentro de los 5
            estricto = arqueo(200, 1653.50, 1450.50, tol=1)
            total = esc(pb1, [(bri + 'B{}'.format(b), 200),
                              (bri + 'C{}'.format(b), 1650.50),
                              (bri + 'D{}'.format(b), 1450.50),
                              (bri + 'B{}'.format(b + 1), 150),
                              (bri + 'C{}'.format(b + 1), 700.25),
                              (bri + 'D{}'.format(b + 1), 620.25)],
                        [('fondo', bri + 'B{}'.format(tot)),
                         ('desc', bri + 'E{}'.format(tot)),
                         ('ver', bri + 'F{}'.format(tot))])

            def temps(valores):
                sets = [(bri + 'B{}'.format(t + i), v)
                        for i, v in enumerate(valores) if v is not None]
                return esc(pb1, sets,
                           [(str(i), bri + 'E{}'.format(t + i))
                            for i in range(len(LIMITES_TEMP))])

            bien = temps([3, -22, 4, 70, 68])
            mal = temps([8, -12, 4, 55, 60])
            # RD-22 · el caso que abría el hallazgo: cámara escarchada.
            escarcha = temps([-5, -22, -6, 70, 68])
            vacio = esc(pb1, [], [('temp', bri + 'E{}'.format(t)),
                                  ('desc', bri + 'E{}'.format(b)),
                                  ('ver', bri + 'F{}'.format(b))])
            fuera['grupo_a_caja_y_temperaturas_bonus01'] = {
                'ref': '{}:Briefing:E{}/F{} (arqueo) y E{} (temperaturas)'
                       .format(FB1, b, b, t),
                'recién descargado': {'conforme': vacio['temp'],
                                      'descuadre': vacio['desc'],
                                      'veredicto': vacio['ver']},
                'caja cuadrada (fondo 200 · contado 1650,50 · Z 1450,50)':
                    {'descuadre': cuadra['desc'], 'veredicto': cuadra['ver']},
                'FALTAN 80 € (el caso que el bloque anterior no veía)':
                    {'descuadre': falta['desc'], 'veredicto': falta['ver']},
                'sobran 10 €': {'descuadre': sobra['desc'],
                                'veredicto': sobra['ver']},
                '+3 € con tolerancia de 5': {'descuadre': dentro['desc'],
                                             'veredicto': dentro['ver']},
                'los mismos +3 € con la tolerancia bajada a 1':
                    {'veredicto': estricto['ver']},
                'TOTAL de las dos cajas': {'fondo': total['fondo'],
                                           'descuadre': total['desc'],
                                           'veredicto': total['ver']},
                'temperaturas correctas (3 · −22 · 4 · 70 · 68)':
                    dict((LIMITES_TEMP[i][0], bien[str(i)])
                         for i in range(len(LIMITES_TEMP))),
                'temperaturas malas (8 · −12 · 4 · 55 · 60)':
                    dict((LIMITES_TEMP[i][0], mal[str(i)])
                         for i in range(len(LIMITES_TEMP))),
                'cámara ESCARCHADA a −5 °C y vitrina fría a −6 °C':
                    dict((LIMITES_TEMP[i][0], escarcha[str(i)])
                         for i in range(len(LIMITES_TEMP))),
                'ok': (vacio['temp'] in ('', None)
                       and vacio['desc'] in ('', None)
                       and vacio['ver'] in ('', None)
                       and abs(cuadra['desc']) < 0.01
                       and 'CUADRA' in str(cuadra['ver'])
                       and abs(falta['desc'] + 80) < 0.01
                       and 'FALTAN' in str(falta['ver'])
                       and abs(sobra['desc'] - 10) < 0.01
                       and 'SOBRAN' in str(sobra['ver'])
                       and 'CUADRA' in str(dentro['ver'])
                       and 'SOBRAN' in str(estricto['ver'])
                       and total['fondo'] == 350
                       # caja 1 cuadra (0) y caja 2 tiene 700,25 − 150
                       # − 620,25 = −70: el total suma los dos.
                       and abs(total['desc'] + 70) < 0.01
                       and 'FALTAN' in str(total['ver'])
                       and all('✓ CONFORME' == bien[str(i)]
                               for i in range(len(LIMITES_TEMP)))
                       and 'FUERA' in str(mal['0'])
                       and 'FUERA' in str(mal['1'])
                       and '✓ CONFORME' == mal['2']
                       and 'FUERA' in str(mal['3'])
                       and 'FUERA' in str(mal['4'])
                       and 'FUERA' in str(escarcha['0'])
                       and 'FUERA' in str(escarcha['2'])),
                'nota': 'el bloque anterior calculaba «Diferencia» = '
                        'recaudación − fondo, que es la recaudación NETA: un '
                        'faltante de 80 € salía exactamente igual que una '
                        'caja cuadrada, y aun así se firmaba como arqueo. Y '
                        'la conformidad de temperatura sólo miraba el límite '
                        'SUPERIOR: una cámara escarchada a −5 °C —producto '
                        'fresco congelado por accidente— salía «✓ CONFORME». '
                        'Ahora hay lectura Z, descuadre con tolerancia en '
                        'celda, y mínimo Y máximo por equipo, con las dos '
                        'filas de mantenimiento en caliente a 65 °C que '
                        'faltaban (DOM-29/RD-04/RD-22/RD-23)'}
    return fuera
