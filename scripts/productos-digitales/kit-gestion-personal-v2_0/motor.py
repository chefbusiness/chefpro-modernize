#!/usr/bin/env python3
"""
motor.py — Motor común del Kit de Gestión de Personal y Turnos v2.0.

Implementa las utilidades TRANSVERSALES del §1 de `kit-gestion-personal-v2-SPEC.md`
sobre los 9 xlsx de `astro-site/public/dl/kit-gestion-personal/`. NO toca ficheros:
recibe un `Workbook` ya cargado y lo modifica en memoria; quien guarda es `main.py`.

Qué hace (§1):
  1.1  LEYENDA ÚNICA como constante — la taxonomía canónica de este kit no son
       categorías de producto sino CÓDIGOS: los 8 de jornada (`M T N P D L V B`,
       con `D` doble nuevo) y los 4 de ausencia (`V B F PE`). El permiso deja de
       ser `P`, que en el 01 significa Partido (DOM-27). De aquí salen la DV, la
       fila 3 de cada rejilla, las líneas de las dos Instrucciones y los colores.
  1.2  Formato condicional REAL (`semaforo()` con `containsText` + reglas
       `expression`). Hoy hay UNA sola regla en las 22 hojas (`04!A7:G65`): el
       «semáforo» que prometen las Instrucciones es el emoji dentro del texto.
  1.3  Capacidad homogénea de 30 empleados como constante (`CAPACIDAD`) y 300
       filas de registro en el 02 (`FILAS_REGISTRO`); los grupos expanden.
  1.4  Catálogo de PARÁMETROS en celda verde con su nota de convenio/CNAE, para
       que ningún grupo se invente un valor por defecto distinto.
  1.5  Guardas: constructores de fórmula (`horas_mod`, `guarda_media`,
       `guarda_resta`, `guarda_division`) + el gate `cacheados_con_error()`.
  1.6  Formatos por CABECERA (hora, fecha, moneda, porcentaje, entero),
       `freeze_panes`, `print_title_rows` / `print_title_cols`, `fitToWidth`.
  1.7  Bio anclada (INSERCIÓN: hoy no la lleva ninguno de los 9), línea
       «Versión 2.0 · agosto 2026 · aichef.pro/kit-gestion-personal ·
       info@aichef.pro», nota de desprotección y metadata OOXML.

Y expone a los grupos: `insertar_columna`, `insertar_fila`, `expandir_filas`,
`marcar_verde`, `sembrar`, `parametro`, `print_setup`, `linea_instrucciones`,
`semaforo`, `_reg`.

════════════════════════════════════════════════════════════════════════════
CENTINELAS — por qué el motor no revienta cuando los grupos aún no existen
════════════════════════════════════════════════════════════════════════════
Los rangos de la SPEC describen el layout FINAL: `01!K` es «Descanso entre
jornadas» sólo DESPUÉS de que `grupo_a` cree esa columna (hoy `K5` está vacía y
la alerta vive en `J`). Si el motor pintara el semáforo en `01!K` a ciegas,
`--solo motor` dejaría reglas de formato sobre una columna vacía.

Por eso cada objetivo lleva CENTINELA POR CABECERA: una columna sólo se toca si
el encabezado de SU bloque ya dice lo que tiene que decir. Es más robusto que el
centinela por celda fija del kit-inventario, porque este kit tiene hojas con
VARIOS bloques (`04!'Checklist Onboarding'` tiene cinco, `07!'Vencimientos'`
dos) y una celda fija no puede validar los dos. Lo que no se aplica se declara
como `pendiente_de_grupo` con fichero:hoja:celda.

ORDEN DEL PIPELINE (`main.py`): grupo.pre → motor.aplicar → grupo.post →
motor.cerrar. Todo lo que depende de rangos vive en `cerrar()`, DESPUÉS de que
los grupos hayan insertado columnas y añadido filas. Regla para los grupos: las
columnas se insertan en `pre()`; las filas se AÑADEN (nunca `insertar_fila` en
medio del bloque de datos) en `post()`.

IDEMPOTENTE: todo es escritura ABSOLUTA y los objetos acumulables —validaciones
y formato condicional— se limpian antes de reescribirse (los nuestros por marca,
los heredados de la v1.1 por SOLAPE de rango). La 2.ª pasada deja el fichero
byte-equivalente salvo timestamps del zip.

Python 3.7 / openpyxl 3.1.3: sin walrus, sin f-strings de depuración (`{x=}`).
pycel NO implementa `COUNTA` ni `MODE`: los sustitutos son `COUNTIF(rango,"<>")`
y `INDEX`/`MATCH` (SPEC, cabecera).
"""
import copy
import re

from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Identidad del producto
# ==========================================================================
PID = 'kit-gestion-personal'
CORTO = 'Kit Gestión de Personal y Turnos'
VERSION = '2.0'

#: §1.7. OJO: el encargo de esta tarea decía «aichef.pro/kit-inventario» —es un
#: arrastre del prompt del kit hermano—. La SPEC §1.7 y el `slug` real
#: (`astro-site/src/data/productos/kits/kit-gestion-personal.ts:25`) dicen
#: `kit-gestion-personal`, y es lo que llevan HOY los 9 ficheros en su línea de
#: versión 1.1. Mandar al cliente a la landing de otro producto sería un enlace
#: roto en el único sitio donde mira cuando quiere la versión nueva.
VERSION_LINE = ('Versión 2.0 · agosto 2026 · aichef.pro/kit-gestion-personal · '
                'info@aichef.pro')
RX_VERSION = re.compile(r'^Versi[óo]n \d+\.\d+ · ')

#: §1.7 — bio anclada. En este kit es una INSERCIÓN, no una sustitución: los 9
#: ficheros terminan hoy en la línea de versión 1.1 y NINGUNO lleva bio, así que
#: `postprocess-transversal.py` (que sólo sustituye patrones conocidos) nunca se
#: la puso. Literal idéntico al de `postprocess-transversal.py:106`.
BIO = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, '
       'en cocina desde los 17 años · johnguerrero.es')

#: §1.6 — la protección va SIN contraseña y hay que decirlo, o el cliente cree
#: que el fichero está cerrado con llave y no lo edita.
NOTA_DESPROTEGER = ('Para editar una celda bloqueada: Revisar → Desproteger '
                    'hoja (no tiene contraseña). Las celdas verdes ya son '
                    'editables sin desproteger nada.')

#: Los MISMOS 9 nombres (§7-bis.1): las 9 claves de descarga de
#: `netlify/functions/get-download-urls.ts:234-243` viajan en emails ya enviados.
FICHEROS = [
    '01-cuadrante-turnos-semanal.xlsx',
    '02-control-horas-extras.xlsx',
    '03-coste-laboral-mensual.xlsx',
    '04-onboarding-nuevo-empleado.xlsx',
    '05-planificacion-vacaciones.xlsx',
    '06-evaluacion-desempeno.xlsx',
    '07-directorio-plantilla.xlsx',
    'BONUS-01-briefing-cambio-turno.xlsx',
    'BONUS-02-calculadora-plantilla-optima.xlsx',
]

# ==========================================================================
# Paleta y formatos
# ==========================================================================
VERDE = 'E8F5E9'         # celda EDITABLE por el cliente (convención de familia)
CAB = '2D2D2D'           # cabecera oscura (la que ya usa la v1.1)
GRIS_BANDA = 'F5F5F5'    # banda alterna por mes en el calendario del 05

CF_VERDE_BG, CF_VERDE_FG = 'C6EFCE', '006100'
CF_AMBAR_BG, CF_AMBAR_FG = 'FFEB9C', '9C6500'
CF_ROJO_BG, CF_ROJO_FG = 'FFC7CE', '9C0006'
CF_AZUL_BG, CF_AZUL_FG = 'BBDEFB', '0D47A1'
CF_NARANJA_BG, CF_NARANJA_FG = 'FFE0B2', 'E65100'
CF_GRIS_BG, CF_GRIS_FG = GRIS_BANDA, '424242'

FMT_HORA = 'hh:mm'
FMT_FECHA = 'dd/mm/yyyy'
FMT_EUR = '#,##0.00 €'
FMT_PCT1 = '0.0%'
FMT_DEC2 = '0.00'
FMT_DEC1 = '0.0'
FMT_ENT = '0'

PIE = '© 2026 AI Chef Pro · aichef.pro'

# ==========================================================================
# §1.1 — LEYENDA ÚNICA (la taxonomía canónica de este kit)
# ==========================================================================
#: Códigos de JORNADA del cuadrante (01). `(código, etiqueta, inicio, fin,
#: horas, color)`. Los horarios salen de la fórmula monstruo de `01!I6`
#: —`IF(B6="M",8,IF(B6="T",8,IF(B6="N",8,IF(B6="P",9,0))))`, medida en
#: `01-cuadrante-turnos-semanal.xlsx:Cuadrante Semanal:I6`— y de las líneas
#: `01!Instrucciones!B16:B20`. `D` (doble, 16 h) es NUEVO: es el código que hace
#: falta para que la alerta de jornada diaria del §2 tenga algo que cazar.
#: `V` y `B` entran en el cuadrante para poder marcar la ausencia donde se
#: planifica el turno, con las MISMAS letras que el 05.
CODIGOS_JORNADA = [
    ('M', 'Mañana',    '07:00', '15:00',  8, 'FFF9C4'),
    # RC-18 · T y V compartían BBDEFB en la MISMA leyenda del 01 (fila 3:
    # `C3='T=Tarde'` y `H3='V=Vacaciones'`, relleno idéntico) mientras
    # `Instrucciones!B8` promete que «cada código lleva su color». El azul
    # BBDEFB se reserva para V —es el que usa el calendario del 05— y la Tarde
    # pasa al cian B2EBF2, libre en las dos paletas.
    ('T', 'Tarde',     '15:00', '23:00',  8, 'B2EBF2'),
    ('N', 'Noche',     '23:00', '07:00',  8, 'E1BEE7'),
    ('P', 'Partido',   '10:00', '23:00',  9, 'FFE0B2'),
    ('D', 'Doble',     '07:00', '23:00', 16, 'FFCDD2'),
    ('L', 'Libre',     '00:00', '00:00',  0, 'C8E6C9'),
    ('V', 'Vacaciones', '00:00', '00:00', 0, 'BBDEFB'),
    ('B', 'Baja',      '00:00', '00:00',  0, 'FFC7CE'),
]

#: Códigos de AUSENCIA del calendario anual (05). El permiso es **PE**, no `P`:
#: en el 01 `P` es Partido y el kit entregaba dos leyendas contradictorias
#: (`05-planificacion-vacaciones.xlsx:Calendario Anual:E3` dice «P=Permiso»
#: mientras `01!Instrucciones!B6` dice «P (partido)»). DOM-27.
CODIGOS_AUSENCIA = [
    ('V',  'Vacaciones', 'BBDEFB'),
    ('B',  'Baja',       'FFC7CE'),
    ('F',  'Festivo',    'C8E6C9'),
    ('PE', 'Permiso',    'FFE0B2'),
]

DV_JORNADA = ','.join(c[0] for c in CODIGOS_JORNADA)          # M,T,N,P,D,L,V,B
#: La coma final deja el valor VACÍO como opción legítima (un día sin marcar no
#: es un error). Es como venía la v1.1 en `05` (`"V,B,F,P,"`).
DV_AUSENCIA = ','.join(c[0] for c in CODIGOS_AUSENCIA) + ','  # V,B,F,PE,

LEYENDA_JORNADA = ('Códigos de jornada: '
                   + ' · '.join('{}={}'.format(c[0], c[1])
                                for c in CODIGOS_JORNADA))
LEYENDA_AUSENCIA = ('Códigos de ausencia: '
                    + ' · '.join('{}={}'.format(c[0], c[1])
                                 for c in CODIGOS_AUSENCIA)
                    + ' (el permiso es PE, no P: en el cuadrante P es Partido)')

#: Dónde va la fila 3 de leyenda (§1.1): `(hoja, fila, códigos)`.
LEYENDA_REJILLA = {
    '01-cuadrante-turnos-semanal.xlsx': [
        ('Cuadrante Semanal', 3, 'jornada'),
        ('Cuadrante Mensual', 3, 'jornada'),
    ],
    '05-planificacion-vacaciones.xlsx': [
        ('Calendario Anual', 3, 'ausencia'),
    ],
}

#: Líneas de Instrucciones que el motor REESCRIBE porque enuncian la leyenda.
#: `(fichero, regex de la línea vieja, texto nuevo)`. El resto de las
#: Instrucciones es trabajo de los grupos (§2/§3/§4).
LINEAS_LEYENDA = [
    ('01-cuadrante-turnos-semanal.xlsx',
     re.compile(r'^▸ Asigna turnos por código'),
     '▸ Asigna turnos por código. ' + LEYENDA_JORNADA + '.'),
    ('01-cuadrante-turnos-semanal.xlsx',
     re.compile(r'^▸ Código de colores'),
     '▸ Cada código lleva su color en la fila 3 de la rejilla: es la misma '
     'leyenda en el cuadrante semanal y en el mensual.'),
    ('05-planificacion-vacaciones.xlsx',
     re.compile(r'^▸ Azul = Vacaciones'),
     '▸ ' + LEYENDA_AUSENCIA + '.'),
]

# ==========================================================================
# §1.3 — capacidad homogénea
# ==========================================================================
#: DOM-32 · hoy la capacidad es 30/20/15/15 según el fichero: el 07 admite 30
#: empleados y el 02 los resume en 15, así que la mitad de la plantilla no cabe
#: en su propio resumen. A partir de aquí, 30 en TODA hoja indexada por empleado.
CAPACIDAD = 30
#: COM-18 · 30 personas generan ~780 registros al mes. 300 filas cubren un mes
#: de 10 personas o una quincena de 20; las Instrucciones lo dicen y ofrecen la
#: salida (hoja por quincena o arrastrar la fórmula).
FILAS_REGISTRO = 300
AVISO_REGISTRO = ('▸ Esta hoja trae {} filas. Con {} personas fichando a diario '
                  'salen unos 780 registros al mes: usa una copia de la hoja '
                  'por quincena, o selecciona la última fila y arrastra hacia '
                  'abajo (las fórmulas se copian solas).'
                  .format(FILAS_REGISTRO, CAPACIDAD))

# ==========================================================================
# §1.4 — catálogo de parámetros en celda verde
# ==========================================================================
#: `nombre → (etiqueta, valor por defecto, formato, nota)`. Ningún grupo escribe
#: un valor por defecto que no esté aquí: el kit entregaba el 0,30 de la SS en
#: 21 fórmulas distintas (`03!Nóminas!D5:D24` y `BONUS-02!Calculadora!B27`) y el
#: 30 de los días de convenio en 30 (`05!Solicitudes!F5:F34`), así que cambiar
#: de convenio obligaba a editar 51 fórmulas a mano.
PARAMETROS = {
    'ss_empresa': (
        'Tipo de SS a cargo de la empresa (%)', 0.33, FMT_PCT1,
        'Ajústalo a tu CNAE, contrato y convenio: un indefinido general ronda '
        'el 33-34 % sumando contingencias comunes, desempleo, AT/EP, FOGASA, '
        'FP y MEI.'),
    'recargo_extra': (
        'Recargo de la hora extra (× la ordinaria) — según tu convenio', 1.25,
        FMT_DEC2,
        'El art. 35.1 ET sólo exige que la hora extra no valga MENOS que la '
        'ordinaria; la cuantía la fija tu convenio o tu contrato, y puede '
        'compensarse con descanso en vez de pagarse.'),
    'tarifa_hora': (
        'Tarifa de la hora ordinaria (€)', 12.0, FMT_EUR,
        'Sácala de la columna «Coste/hora» del 03: es el coste real por hora, '
        'no el bruto del convenio.'),
    'limite_extra_anual': (
        'Límite anual de horas extra', 80, FMT_ENT,
        'Art. 35.2 ET. NO computan en el tope las compensadas con descanso '
        'dentro de los 4 meses siguientes ni las de fuerza mayor.'),
    'dias_convenio': (
        'Días de vacaciones por convenio (naturales)', 30, FMT_ENT,
        'Art. 38 ET: 30 días naturales al año. Muchos convenios de hostelería '
        'lo expresan como 22 días laborables — si es tu caso, escribe 22 y '
        'cuenta sólo días de trabajo en las solicitudes.'),
    'jornada_semanal': (
        'Jornada contratada (h/semana)', 40, FMT_ENT,
        'Es la de un contrato a tiempo completo. Los parciales llevan la suya '
        'en su fila: por eso la columna es verde y editable persona a persona.'),
    'jornada_diaria_max': (
        'Máx. horas de jornada ordinaria diaria', 9, FMT_ENT,
        'Art. 34.3 ET. Se puede superar por distribución irregular pactada, '
        'pero respetando el descanso entre jornadas.'),
    'descanso_min': (
        'Descanso mínimo entre jornadas (h)', 12, FMT_ENT,
        'Art. 34.3 ET: 12 horas. Las 11 h que se leen por ahí son el mínimo de '
        'la Directiva 2003/88/CE, que en España está mejorado.'),
    'horas_por_servicio': (
        'Horas efectivas por servicio', 4, FMT_ENT,
        'Lo que dura de verdad un servicio con su montaje y su cierre. Súbelo '
        'si tu turno es de 5 h o si encadenas montaje y desmontaje.'),
    'factor_cobertura': (
        'Factor de cobertura (vacaciones, bajas y descansos)', 1.15, FMT_DEC2,
        'Un 15 % adicional sobre el personal en sala para cubrir libranzas, '
        'vacaciones y bajas. En temporada alta súbelo a 1,20.'),
    'dias_apertura': (
        'Días de apertura / semana', 6, FMT_ENT,
        'Entra en el cálculo por las horas semanales: abrir 7 días en vez de 6 '
        'sube la plantilla necesaria, no sólo el turno.'),
    'umbral_semaforo': (
        'Umbral del semáforo (ratio objetivo / aceptable)', None, FMT_PCT1,
        'Sale por VLOOKUP del tipo de negocio que elijas: no hay un 30 % bueno '
        'para todos. Un fine dining al 38 % está donde debe y un fast casual '
        'al 38 % está en pérdidas.'),
}

# ==========================================================================
# Registro de fórmulas (main.py verifica una a una que quedaron con cache)
# ==========================================================================
REGISTRO = []


def _reg(ws, coord, formula):
    """Anota una fórmula escrita para que `main.py` la verifique con
    `data_only`. Los grupos DEBEN llamarlo por cada fórmula que escriban."""
    REGISTRO.append((ws.title, coord, formula))
    return formula


# ==========================================================================
# §1.5 — constructores de fórmula con guarda
# ==========================================================================
def horas_mod(col_ent, col_sal, fila, col_pausa=None):
    """Horas trabajadas con CRUCE DE MEDIANOCHE y pausa (§7-bis.2, DOM-03).

    `MOD` sí evalúa con horas —medido con pycel el 2026-08-23: `MOD(D5-C5,1)*24`
    con 23:00→07:00 da `7.999999999999998`—, así que el `ROUND(...,2)` NO es
    cosmético: sin él `inject_cache` graba ese ruido en el fichero que se
    descarga y el cliente ve «7,999999999999998 h» en el visor del móvil.

    La v1.1 hacía `(D5-C5)*24` (`02!Registro Horas!E5`), que con un turno de
    noche da **−16 h**: el fichaje del turno de noche entraba en negativo.
    """
    pausa = '0' if col_pausa is None else \
        'IF(${p}{f}="",0,${p}{f})'.format(p=col_pausa, f=fila)
    return ('=IF(OR(${c}{f}="",${s}{f}=""),"",'
            'ROUND(MOD(${s}{f}-${c}{f},1)*24-{p},2))'
            .format(c=col_ent, s=col_sal, f=fila, p=pausa))


def horas_alternativa(col_ent, col_sal, fila):
    """Alternativa documentada al `MOD` (§7-bis.2): da 8,0 exacto sin ROUND.
    Se deja escrita en las Instrucciones para quien use una hoja de cálculo
    donde `MOD` con horas no se comporte."""
    return ('=IF(${c}{f}="","",IF(${s}{f}<${c}{f},'
            '${s}{f}+1-${c}{f},${s}{f}-${c}{f})*24)'
            .format(c=col_ent, s=col_sal, f=fila))


def guarda_media(rango, decimales=2):
    """§1.5 — media que NO devuelve `#¡DIV/0!` con el rango vacío.

    `06!Ficha Evaluación!C22` es hoy `=AVERAGE(C12:C21)` y la ficha recién
    descargada enseña `#¡DIV/0!` en un documento que se FIRMA (DOM-02/COM-04).
    pycel no implementa `COUNTA`, pero `COUNT` sí, y aquí lo que se cuenta son
    números: es el sustituto correcto.
    """
    return ('=IF(COUNT({r})=0,"",ROUND(AVERAGE({r}),{d}))'
            .format(r=rango, d=decimales))


def guarda_resta(a, b, expresion=None):
    """§1.5 — resta/producto que se calla si falta cualquiera de los dos datos.
    Sin esto, `02!Registro Horas!G5` declara HORA EXTRA la jornada entera en
    cuanto la columna «H. Contratadas» está vacía: 8 h extra × 21 €/h = 168 €
    por día y empleado que el cliente no debe (DOM-11/TEC-09)."""
    if expresion is None:
        expresion = '{}-{}'.format(a, b)
    return '=IF(OR({a}="",{b}=""),"",{e})'.format(a=a, b=b, e=expresion)


def guarda_division(num, den, decimales=4):
    """§1.5 — división con `IFERROR` y guarda de denominador cero."""
    return ('=IFERROR(IF(OR({d}="",{d}=0),"",ROUND({n}/{d},{k})),"")'
            .format(n=num, d=den, k=decimales))


def contar_no_vacias(rango):
    """Sustituto de `COUNTA`, que pycel NO implementa (SPEC, cabecera)."""
    return 'COUNTIF({},"<>")'.format(rango)


# ==========================================================================
# Bloques de datos: (hoja, fila_cabecera, primera_fila, ULTIMA_HOY, ULTIMA_V2)
# ==========================================================================
#: Hacen falta las DOS últimas filas, no basta con `ws.max_row`: por debajo del
#: bloque hay COLA (el «TOTAL HORAS EQUIPO» de `01!Cuadrante Semanal!A21`, el
#: «TOTALES» de `02!Resumen Mensual!A22`, el pie «© 2026 AI Chef Pro»). Regla:
#: si `ws.max_row` ya supera `ULTIMA_V2`, el grupo expandió el bloque y se usa
#: `ULTIMA_V2`; si no, sigue siendo el de la v1.1 y se usa `ULTIMA_HOY`.
#:
#: Una hoja puede tener VARIOS bloques (`04` cinco tramos, `07!Vencimientos`
#: dos, `BONUS-01!Briefing` cinco): por eso `_rangos_filas()` devuelve una LISTA.
#:
#: Se omiten a propósito:
#:   · `01!'Cuadrante Mensual'` — hoy es una rejilla MUERTA de 4 bloques sin una
#:     sola fórmula ni DV (`01-cuadrante-turnos-semanal.xlsx:Cuadrante
#:     Mensual:A3..A61`); `grupo_a` la reconstruye entera a 5 semanas (COM-12).
#:   · `03!'Previsión por Servicio'` y `BONUS-02!'Calculadora'` — son hojas de
#:     PARÁMETROS en vertical (rótulo en A, valor en B), no tablas.
BLOQUES = {
    '01-cuadrante-turnos-semanal.xlsx': [
        ('Cuadrante Semanal', 5, 6, 20, 5 + CAPACIDAD),          # 6..35
    ],
    '02-control-horas-extras.xlsx': [
        ('Registro Horas', 4, 5, 54, 4 + FILAS_REGISTRO),        # 5..304
        ('Resumen Mensual', 5, 6, 20, 5 + CAPACIDAD),            # 6..35
    ],
    '03-coste-laboral-mensual.xlsx': [
        ('Nóminas', 4, 5, 24, 4 + CAPACIDAD),                    # 5..34
        # tabla de referencia de ratios por tipo de negocio (A13:C19 hoy; el
        # grupo_b le añade las dos columnas numéricas del VLOOKUP)
        # RD-21 · el bloque de coste gana cuatro filas, así que la tabla de
        # referencia baja de la 13 a la 17. RC-11 · y pasa de 6 a 10 tipos.
        ('Ratio Coste Laboral', 17, 18, 27, 27),
    ],
    '04-onboarding-nuevo-empleado.xlsx': [
        # los 5 tramos se DETECTAN (ver `secciones_04`): sus filas se mueven en
        # cuanto grupo_c añade las 3 tareas que faltan (DOM-15), y una tabla
        # fija aquí quedaría desfasada en la misma pasada que las añade.
    ],
    '05-planificacion-vacaciones.xlsx': [
        ('Calendario Anual', 5, 6, 20, 5 + CAPACIDAD),           # 6..35
        # RD-09/RT-03 · «Solicitudes» se indexa por PETICIÓN, no por
        # empleado: 30 filas para 30 personas era una solicitud por persona y
        # año. Cinco periodos por persona.
        ('Solicitudes', 4, 5, 34, 4 + 5 * CAPACIDAD),            # 5..154
        ('Cobertura', 5, 6, 8, 8),                               # turnos
        ('Cobertura', 11, 12, 27, 11 + CAPACIDAD),               # sustituciones
    ],
    '06-evaluacion-desempeno.xlsx': [
        ('Ficha Evaluación', 11, 12, 21, 21),                    # competencias
        ('Histórico', 4, 5, 19, 4 + CAPACIDAD),                  # 5..34
    ],
    '07-directorio-plantilla.xlsx': [
        ('Plantilla', 4, 5, 34, 4 + CAPACIDAD),                  # 5..34
        ('Vencimientos', 6, 7, 21, 6 + CAPACIDAD),               # contratos
        ('Vencimientos', 25, 26, 35, 35),                        # carnets (v1)
    ],
    'BONUS-01-briefing-cambio-turno.xlsx': [
        ('Briefing', 8, 9, 14, 14),        # reservas / VIPs
        ('Briefing', 16, 17, 22, 22),      # incidencias
        ('Briefing', 24, 25, 30, 30),      # tareas pendientes
        ('Briefing', 32, 33, 38, 38),      # stock bajo
        ('Briefing', 40, 41, 45, 45),      # personal
    ],
    'BONUS-02-calculadora-plantilla-optima.xlsx': [
        ('Ratios por Tipo', 4, 5, 14, 14),
    ],
}

#: Hojas de REFERENCIA o de PARÁMETROS: el verde no se deduce solo (lo pondría
#: sobre la base normativa o sobre el rótulo de un parámetro). Los grupos marcan
#: ahí las celdas concretas con `marcar_verde()`.
SIN_VERDE_AUTO = frozenset({
    'Ratio Coste Laboral',   # 03: tabla de ratios de referencia + parámetros
    'Ratios por Tipo',       # BONUS-02: tabla de referencia
    'Previsión por Servicio',  # 03: parámetros en vertical
    'Calculadora',           # BONUS-02: parámetros en vertical
})

#: Hojas que hay que proteger AUNQUE no tengan ni una celda verde: son 100 %
#: fórmulas y un clic borra la cadena entera. `07!Vencimientos` NO está aquí
#: porque su segundo bloque (carnets, A26:C35) sí se escribe a mano.
PROTEGER_SIN_VERDE = frozenset()

#: Cabeceras que NUNCA llevan verde aunque su columna no tenga fórmulas: son
#: contenido IMPRESO (la lista de competencias, las 50 tareas del onboarding,
#: la taxonomía de la tabla de ratios), no casillas de entrada.
NO_VERDE_CABECERAS = frozenset({
    '#', 'tarea', 'categoría', 'categoria', 'competencia', 'turno',
    'tipo de negocio', 'parámetro', 'parametro', 'ratio objetivo',
    'rango aceptable', 'covers/cocinero', 'covers/camarero', 'covers/barman',
    'ratio coste laboral', 'código', 'codigo', 'descripción', 'descripcion',
})

# ==========================================================================
# §1.1 — dónde va la validación de datos
# ==========================================================================
#: `(hoja, columna_desde, columna_hasta, texto_esperado_en_la_cabecera, códigos)`.
#: El centinela es la CABECERA DEL BLOQUE, no una celda fija: `01!B5` dice
#: «Lunes» y `05!B5` dice «Ene» hoy y dirá «1» (nº de semana) cuando grupo_c
#: reconstruya el calendario, así que la comprobación se hace sobre la fila de
#: cabecera que corresponda a cada bloque.
DV_CODIGOS = {
    '01-cuadrante-turnos-semanal.xlsx': [
        ('Cuadrante Semanal', 'B', 'H', 'Lunes', 'jornada'),
    ],
    '05-planificacion-vacaciones.xlsx': [
        # hasta que grupo_c reconstruya el calendario por semanas, las columnas
        # son los 12 meses (B..M). Después serán B..BB. El rango se recorta a
        # `ws.max_column - 1` (la última columna es el total «Días Usados»).
        ('Calendario Anual', 'B', None, None, 'ausencia'),
    ],
}

#: DV de LISTA sueltas que no son códigos: `(hoja, columna, texto_cabecera,
#: valores, título, prompt)`. Son las que la SPEC pide ampliar o crear.
DV_LISTA = {
    '02-control-horas-extras.xlsx': [
        ('Registro Horas', 'H', 'Tipo Extra',
         ['Voluntaria', 'Obligatoria', 'Fuerza mayor',
          'Compensada con descanso'],
         'Tipo de hora extra no válido',
         'Art. 35.2 ET: las de FUERZA MAYOR y las COMPENSADAS CON DESCANSO '
         'dentro de los 4 meses siguientes NO computan en el tope de 80 h/año. '
         'El Resumen Mensual las descuenta por este texto.'),
    ],
    '05-planificacion-vacaciones.xlsx': [
        ('Solicitudes', 'G', 'Estado',
         ['Pendiente', 'Aprobado', 'Denegado'],
         'Estado no válido',
         'El saldo de «Saldo Vacaciones» sólo suma los días de las solicitudes '
         'APROBADAS, así que el texto tiene que salir de esta lista.'),
    ],
    '07-directorio-plantilla.xlsx': [
        ('Plantilla', 'D', 'Tipo Contrato',
         ['Indefinido', 'Temporal', 'Prácticas', 'Formación',
          'Fijo-discontinuo'],
         'Tipo de contrato no válido', 'Modalidades vigentes tras el RDL '
         '32/2021.'),
        ('Plantilla', 'G', 'Jornada', ['Completa', 'Parcial'],
         'Jornada no válida', 'Una jornada parcial no puede hacer horas extra '
         '(art. 12.4.c ET): sólo horas complementarias pactadas.'),
    ],
    'BONUS-01-briefing-cambio-turno.xlsx': [
        ('Briefing', 'C', 'Gravedad', ['Alta', 'Media', 'Baja'],
         'Gravedad no válida',
         'TEC-28 · sin lista, cada turno escribe una palabra distinta y el '
         'briefing deja de ser comparable de un día para otro.'),
        ('Briefing', 'A', 'Prioridad', ['Urgente', 'Alta', 'Normal'],
         'Prioridad no válida', 'Lo que el turno entrante mira primero.'),
    ],
}

#: Marca de las DV del motor, para poder limpiarlas antes de reescribirlas.
MARCA_DV = 'kitgp-v2'

# ==========================================================================
# §1.2 — formato condicional
# ==========================================================================
#: Vocabularios: `(subcadena, color)`, evaluados EN ORDEN con `stopIfTrue`, así
#: que lo más grave gana. `containsText` usa `SEARCH`, que no distingue
#: mayúsculas y no se pelea con el emoji de delante.
VOC_ALERTA = [('⛔', 'rojo'), ('EXCESO', 'rojo'), ('⚠', 'ambar'),
              ('OK', 'verde')]
VOC_LIMITE = [('EXCEDE', 'rojo'), ('cerca del límite', 'ambar'),
              ('dentro', 'verde'), ('OK', 'verde')]
VOC_RATIO = [('ACCIÓN CORRECTIVA', 'rojo'), ('VIGILAR', 'ambar'),
             ('EXCELENTE', 'verde')]
#: RD-24 · «puntúa al menos» es el aviso de datos insuficientes de la ficha
#: (06!C23): un nivel emitido con UNA competencia puntuada no es un nivel.
#: Va en ámbar y no comparte ninguna palabra con los cinco veredictos, así
#: que el orden entre ellos no importa.
VOC_NIVEL = [('DEFICIENTE', 'rojo'), ('MEJORABLE', 'ambar'),
             ('ADECUADO', 'ambar'), ('EXCELENTE', 'verde'),
             ('BUENO', 'verde'), ('puntúa al menos', 'ambar')]
VOC_VENCIMIENTO = [('VENCIDO', 'rojo'), ('URGENTE', 'rojo'), ('🔴', 'rojo'),
                   ('🟡', 'ambar'), ('PRONTO', 'ambar'), ('🟢', 'verde'),
                   ('OK', 'verde')]
VOC_COBERTURA = [('TEMP. ALTA', 'rojo'), ('EXCESO', 'ambar')]
VOC_CONFORME = [('FUERA DE RANGO', 'rojo'), ('CONFORME', 'verde')]
#: RD-04 · veredicto del arqueo de caja del BONUS-01. Un FALTANTE es rojo
#: (dinero que no está) y un SOBRANTE es ámbar: también es un descuadre —
#: casi siempre un cobro no registrado— pero no un agujero de caja.
VOC_CAJA = [('FALTAN', 'rojo'), ('SOBRAN', 'ambar'),
            ('CUADRA', 'verde')]

#: Semáforos por COLUMNA, con centinela de cabecera:
#: `(hoja, columna, texto_esperado_en_cabecera, vocabulario)`.
#: Se declaran a la vez la columna de HOY y la de la v2.0. Como `cerrar()` corre
#: DESPUÉS de los grupos, el centinela decide sola cuál existe: con `--solo
#: motor` se colorea la alerta que el fichero ya tiene; con `grupo_a` cargado,
#: las cuatro nuevas. Ninguna regla queda sobre una columna que cambió de
#: significado, porque `_limpiar_cf` borra por rango antes de reescribir.
CF_COLUMNA = {
    '01-cuadrante-turnos-semanal.xlsx': [
        ('Cuadrante Semanal', 'J', 'Alerta', VOC_ALERTA),        # v1.1
        ('Cuadrante Semanal', 'K', 'Descanso entre', VOC_ALERTA),
        ('Cuadrante Semanal', 'L', 'Descanso semanal', VOC_ALERTA),
        ('Cuadrante Semanal', 'M', 'Jornada semanal', VOC_ALERTA),
        ('Cuadrante Semanal', 'N', 'Jornada diaria', VOC_ALERTA),
    ],
    '02-control-horas-extras.xlsx': [
        ('Resumen Mensual', 'D', 'Dentro Límite', VOC_LIMITE),   # v1.1
        ('Resumen Mensual', 'F', 'Límite anual', VOC_LIMITE),    # v2.0
    ],
    '07-directorio-plantilla.xlsx': [
        ('Vencimientos', 'C', 'Alerta', VOC_VENCIMIENTO),
        ('Vencimientos', 'E', 'Alerta', VOC_VENCIMIENTO),
        ('Vencimientos', 'G', 'Alerta', VOC_VENCIMIENTO),
        ('Vencimientos', 'I', 'Alerta', VOC_VENCIMIENTO),
    ],
    'BONUS-01-briefing-cambio-turno.xlsx': [
        ('Briefing', 'D', 'Conforme', VOC_CONFORME),
    ],
}

#: Semáforos sobre una CELDA o un rango fijo que no cuelga de ningún bloque:
#: `(hoja, rango, celda_centinela, texto_centinela, vocabulario)`.
CF_CELDA = {
    '03-coste-laboral-mensual.xlsx': [
        ('Ratio Coste Laboral', 'B9:B9', 'A9', 'Semáforo', VOC_RATIO),
    ],
    '06-evaluacion-desempeno.xlsx': [
        ('Ficha Evaluación', 'C23:C23', 'B23', 'NIVEL', VOC_NIVEL),
    ],
    '05-planificacion-vacaciones.xlsx': [
        # fila 38 del calendario: alerta de cobertura (§4). Sólo existe cuando
        # grupo_c la construye; el centinela es el rótulo de `A38`.
        # RD-12/RD-13/RT-11 · la fila de ausencias se desdobla (rejilla ·
        # solicitudes · la mayor), así que la alerta baja de la 38 a la 40.
        ('Calendario Anual', 'B40:BB40', 'A40', 'Cobertura', VOC_COBERTURA),
    ],
}

#: Reglas `expression`: `(hoja, rango, celda_centinela, texto, fórmula, color,
#: stop)`. `{f}` se sustituye por la primera fila del rango.
#:
#: La banda gris por MES del calendario del 05 (§1.2) va con `stopIfTrue=False`
#: y DESPUÉS de las reglas de código, o taparía la marca de vacaciones. La
#: evalúa Excel, no pycel: `ISODD(MONTH(...))` sobre la fila de fechas.
CF_EXPRESION = {
    '05-planificacion-vacaciones.xlsx': [
        ('Calendario Anual', 'B6:BB35', 'A4', 'Semana',
         '=ISODD(MONTH(B$5))', 'gris', False),
    ],
}

# ==========================================================================
# §1.6 — formatos deducidos de la CABECERA
# ==========================================================================
#: Se decide por el TEXTO del encabezado, no por la letra de la columna: así
#: sobrevive a que un grupo inserte «Pausa (h)» delante de «Horas trabajadas».
#: El orden importa — se para en la primera coincidencia, y por eso «hora
#: entrada» va antes que el genérico «horas».
FORMATO_POR_CABECERA = [
    (FMT_HORA, ('hora entrada', 'hora salida', 'hora inicio', 'hora fin',
                'entrada', 'salida')),
    (FMT_FECHA, ('fecha', 'caducidad', 'vencimiento', 'fin de contrato',
                 'fin periodo', 'alta')),
    (FMT_PCT1, ('%', 'porcentaje', 'ratio coste')),
    (FMT_EUR, ('€', 'eur', 'salario', 'coste', 'tarifa', 'importe', 'bruto',
               'nómina', 'nomina', 'recaudación', 'recaudacion', 'fondo')),
    (FMT_ENT, ('días', 'dias', 'pagas', 'nº de', 'n° de', 'cubiertos',
               'covers', 'semanas', 'mínimo personal', 'minimo personal')),
    (FMT_DEC2, ('horas', 'h. contratadas', 'h. extra', 'h. acumuladas',
                'jornada', 'pausa')),
    (FMT_DEC1, ('puntuación', 'puntuacion', 'media', 'q1', 'q2', 'q3', 'q4')),
]

#: Rangos de formato que NINGUNA cabecera delata: son parámetros sueltos en
#: vertical (rótulo en A, valor en B). `(hoja, rango, formato, celda_centinela,
#: texto_centinela)`.
#:
#: §1.6 los cita uno a uno porque hoy el dinero de ENTRADA va en `General`:
#: `02!Resumen Mensual!B3` vale 12 y se lee «12», no «12,00 €».
RANGOS_FORMATO = {
    '02-control-horas-extras.xlsx': [
        ('Resumen Mensual', 'B3:B3', FMT_EUR, 'A3', 'Tarifa'),
    ],
    '03-coste-laboral-mensual.xlsx': [
        ('Ratio Coste Laboral', 'B4:B5', FMT_EUR, 'A4', 'Ventas'),
        ('Previsión por Servicio', 'B22:B22', FMT_EUR, 'A22', 'Coste'),
    ],
    'BONUS-02-calculadora-plantilla-optima.xlsx': [
        ('Calculadora', 'B9:B9', FMT_EUR, 'A9', 'Salario'),
    ],
}

# ==========================================================================
# §1.6 — presentación: freeze, títulos de impresión, orientación
# ==========================================================================
#: `(hoja, freeze_panes, print_title_rows, print_title_cols, landscape,
#:   fit_to_width)`. `None` = no tocar / valor por defecto.
#:
#: Las 5 hojas sin `freeze_panes` medidas hoy: `07!Plantilla`,
#: `01!Cuadrante Mensual`, `03!Ratio Coste Laboral`,
#: `03!Previsión por Servicio` y `BONUS-02!Calculadora` (§1.6, TEC-22).
PRESENTACION = {
    '01-cuadrante-turnos-semanal.xlsx': [
        ('Cuadrante Semanal', 'B6', '$5:$5', '$A:$A', True, 1),
        # RT-17 · el título repetido era '$4:$4', la cabecera del bloque
        # SEMANA 1, y la hoja tiene SEIS bloques con su propia cabecera (filas
        # 4, 37, 70, 103, 136 y 169) en 202 filas: al imprimir se repetía en
        # todas las páginas la cabecera de la primera semana, y como la fila 3
        # («SEMANA 1» + leyenda) quedaba fuera, las páginas de las semanas 2 a
        # 5 salían sin decir a qué semana pertenecen. Se repite el TÍTULO del
        # documento ('$1:$2'), que vale para los seis bloques, y `grupo_a`
        # mete un salto de página antes de cada uno para que cada semana
        # empiece en su propia hoja con su cabecera visible.
        ('Cuadrante Mensual', 'A5', '$1:$2', '$A:$A', True, 1),
    ],
    '02-control-horas-extras.xlsx': [
        ('Registro Horas', 'A5', '$4:$4', None, True, 1),
        ('Resumen Mensual', 'A6', '$5:$5', None, True, 1),
    ],
    '03-coste-laboral-mensual.xlsx': [
        ('Nóminas', 'A5', '$4:$4', None, True, 1),
        # RT-22 · el freeze estaba en A14 e inmovilizaba 13 de las 22 filas de
        # la hoja: la zona desplazable se quedaba en las nueve de la tabla de
        # referencia. Se baja a A13 (sólo la cabecera de esa tabla queda fija),
        # que es lo que hacen 'Previsión' (A4 sobre 26 filas) y 'Nóminas' (A5
        # sobre 37).
        # RD-21 · el bloque de coste gana cuatro filas (nóminas + horas
        # extra + ETT + otros = total), así que la cabecera de la tabla de
        # referencia baja de la 13 a la 17. RT-22 · el freeze estaba en A14 e
        # inmovilizaba 13 de las 22 filas de la hoja: la zona desplazable se
        # quedaba en las nueve de la tabla. Ahora fija lo que hay que tener a
        # la vista mientras se baja a la tabla (el ratio y su semáforo) y deja
        # el resto desplazable.
        ('Ratio Coste Laboral', 'A15', '$17:$17', None, False, 1),
        ('Previsión por Servicio', 'A4', None, None, False, 1),
    ],
    '04-onboarding-nuevo-empleado.xlsx': [
        ('Checklist Onboarding', 'A7', '$6:$6', '$A:$B', True, 1),
    ],
    '05-planificacion-vacaciones.xlsx': [
        # 53 semanas: sin `print_title_cols` la página 2 en adelante se imprime
        # sin los nombres y no se sabe de quién es cada fila.
        ('Calendario Anual', 'B6', '$4:$5', '$A:$A', True, 1),
        ('Solicitudes', 'A5', '$4:$4', None, True, 1),
        ('Cobertura', 'A6', '$5:$5', None, False, 1),
    ],
    '06-evaluacion-desempeno.xlsx': [
        ('Ficha Evaluación', None, None, None, False, 1),
        ('Histórico', 'A5', '$4:$4', None, True, 1),
    ],
    '07-directorio-plantilla.xlsx': [
        # §7.6 · 21 columnas: se acepta que se imprima en DOS páginas A4 con el
        # nombre repetido, antes que partir la hoja en dos bloques verticales.
        ('Plantilla', 'B5', '$4:$4', '$A:$A', True, 2),
        ('Vencimientos', 'A7', '$6:$6', '$A:$A', True, 1),
    ],
    'BONUS-01-briefing-cambio-turno.xlsx': [
        ('Briefing', None, None, None, False, 1),
    ],
    'BONUS-02-calculadora-plantilla-optima.xlsx': [
        ('Calculadora', 'A5', None, None, False, 1),
        ('Ratios por Tipo', 'A5', '$4:$4', None, False, 1),
    ],
}

# ==========================================================================
# 04 — colores de sección (§1.2) y colisión con el verde de edición
# ==========================================================================
#: `04!Instrucciones!B11:B15` anuncia cinco colores y la hoja no pinta ninguna
#: cabecera de sección: el color vive sólo dentro de la columna «Categoría».
#:
#: ⚠ Y ahí hay una COLISIÓN medida: `04-onboarding-nuevo-empleado.xlsx:Checklist
#: Onboarding:C31` … `:C39` están rellenas de **E8F5E9**, que es exactamente el
#: verde de edición de la familia. Cualquier `es_verde()` las tomaría por
#: casillas editables y `proteger()` las dejaría desbloqueadas «porque son
#: verdes». Se repintan a `C8E6C9` —sigue siendo el verde que anuncian las
#: Instrucciones— para que `E8F5E9` signifique EDITABLE y sólo eso en los 9
#: ficheros. `(subcadena del título de sección, color)`.
SECCIONES_04 = [
    ('DOCUMENTACIÓN', 'E3F2FD'),
    ('FORMACIÓN OBLIGATORIA', 'FFF3E0'),
    ('EQUIPAMIENTO', 'C8E6C9'),
    ('FORMACIÓN OPERATIVA', 'F3E5F5'),
    ('PERIODO DE PRUEBA', 'FFEBEE'),
]

# ==========================================================================
# Gate auxiliar: citas legales que la v1.1 imprime mal
# ==========================================================================
#: `(regex, motivo)`. NO se corrigen aquí —el texto de cada fichero es trabajo
#: de su grupo—, pero el motor los CUENTA para que `--solo motor` diga cuántas
#: afirmaciones falsas siguen vivas en el kit. Todas están medidas.
CITAS_OBSOLETAS = [
    (re.compile(r'11\s*h.*descanso|descanso.*11\s*h', re.I),
     'descanso entre jornadas: son 12 h (art. 34.3 ET); las 11 h son el '
     'mínimo de la Directiva 2003/88/CE'),
    (re.compile(r'×\s*1[.,]75|x\s*1[.,]75', re.I),
     'el ×1,75 se presenta como «Legislación española» y no lo es: art. 35.1 '
     'ET sólo exige que no valga menos que la ordinaria'),
    (re.compile(r'×\s*2[.,]0|x\s*2[.,]0', re.I),
     'el ×2,0 se presenta como «Legislación española» y no lo es'),
    (re.compile(r'~?\s*30\s*%.*[Ss]eguridad [Ss]ocial|[Ss]eguridad [Ss]ocial'
                r'.*30\s*%'),
     'la SS a cargo de la empresa ronda el 33-34 %, no el 30 %, y tiene que '
     'ser un parámetro editable por CNAE'),
    (re.compile(r'al[ée]rgenos propios', re.I),
     'dato de salud, categoría especial del art. 9 RGPD: se elimina (DOM-13)'),
]


# ==========================================================================
# Utilidades de estilo
# ==========================================================================
def _relleno(cel):
    f = cel.fill
    if f is None or f.fill_type != 'solid' or f.fgColor is None:
        return None
    rgb = f.fgColor.rgb
    return rgb.upper() if isinstance(rgb, str) else None


def es_verde(cel):
    """True si la celda lleva el verde de edición del kit."""
    r = _relleno(cel)
    return bool(r) and r.endswith(VERDE)


def marcar_verde(ws, rango):
    """Pinta de verde (= editable) un rango `A1:B9` o una celda `A1`."""
    n = 0
    filas = ws[rango]
    # `ws['C4']` devuelve un Cell suelto (no iterable) y `ws['C4:D9']` una tupla
    # de tuplas: sin esto, marcar UNA celda revienta con «'Cell' object is not
    # iterable» pese a lo que promete la docstring.
    if not isinstance(filas, tuple):
        filas = ((filas,),)
    for fila in filas:
        celdas = fila if isinstance(fila, tuple) else (fila,)
        for c in celdas:
            if not es_verde(c):
                c.fill = PatternFill('solid', fgColor=VERDE)
            c.protection = Protection(locked=False)
            n += 1
    return n


def sembrar(ws, fila, valores, col0=1, marca=None):
    """Escribe una fila de ejemplo. `None` deja la celda como esté (útil para no
    pisar una fórmula). `marca` se concatena a la primera celda de texto."""
    escritas = 0
    for i, v in enumerate(valores):
        if v is None:
            continue
        cel = ws.cell(row=fila, column=col0 + i)
        if marca and isinstance(v, str) and escritas == 0:
            v = v + ' ' + marca
        cel.value = v
        escritas += 1
    return escritas


def parametro(ws, fila, clave, col_rotulo=1, col_valor=2, valor=None,
              nota=True):
    """§1.4 — escribe un parámetro del catálogo: rótulo, celda VERDE con el
    valor por defecto, formato y —si `nota`— la explicación de convenio/CNAE
    como mensaje de entrada de una DV informativa.

    Devuelve la coordenada de la celda de valor, que es lo que las fórmulas
    tienen que referenciar: NUNCA el literal dentro de la fórmula (§1.4).
    """
    etiqueta, defecto, fmt, texto = PARAMETROS[clave]
    ws.cell(row=fila, column=col_rotulo, value=etiqueta + ':')
    cel = ws.cell(row=fila, column=col_valor)
    cel.value = defecto if valor is None else valor
    cel.number_format = fmt
    cel.fill = PatternFill('solid', fgColor=VERDE)
    cel.protection = Protection(locked=False)
    if nota and texto:
        dv = DataValidation(type=None, showInputMessage=True,
                            promptTitle='{} · {}'.format(MARCA_DV, clave),
                            prompt=texto)
        ws.add_data_validation(dv)
        dv.add(cel.coordinate)
    return cel.coordinate


# ==========================================================================
# Utilidades genéricas (probadas en kit-pasteleria/escandallos/inventario v2)
# ==========================================================================
RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _traducir_formula(valor, idx, eje):
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        ci = column_index_from_string(col)
        fi = int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        return '{}{}{}{}'.format(d1, col, d2, fila)

    return RX_REF.sub(_sub, valor)


def _rangos_dv(ws):
    return [({k: getattr(dv, k, None) for k in CAMPOS_DV},
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**{k: v for k, v in attrs.items() if v is not None})
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje) if idx else r)


def _desplazar_rango(ref, idx, eje):
    partes = ref.split(':')
    fuera = []
    for p in partes:
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        fuera.append('{}{}{}{}'.format(d1, col, d2, fila))
    return ':'.join(fuera)


def insertar_columna(ws, idx):
    """Inserta una columna en `idx` manteniendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos. Los grupos la llaman desde
    `pre()`, ANTES de que el motor fije rangos."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos = dict((k, v.width) for k, v in ws.column_dimensions.items()
                  if v.width)

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra, ancho in sorted(anchos.items(),
                               key=lambda kv: -column_index_from_string(kv[0])):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = ancho


def insertar_fila(ws, idx):
    """Equivalente por filas de `insertar_columna`."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = dict((k, v.height) for k, v in ws.row_dimensions.items()
                   if v.height)

    for fila in range(max_r, idx - 1, -1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila + 1, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'fila'))
    _restaurar_dv(ws, dvs, idx, 'fila')

    for fila, alto in sorted(alturas.items(), reverse=True):
        if fila >= idx:
            ws.row_dimensions[fila + 1].height = alto


def expandir_filas(ws, r_ult, r_destino, cola=(), numerar=None):
    """§1.3 — añade filas AL FINAL del bloque replicando la última (estilo y
    fórmulas traducidas) y baja la «cola» (totales, pie) `r_destino - r_ult`.

    CENTINELA de idempotencia: si `r_destino` ya tiene contenido, no hace nada.
    """
    if r_destino <= r_ult:
        return 0
    max_c = ws.max_column
    if any(ws.cell(row=r_destino, column=c).value is not None
           for c in range(1, max_c + 1)):
        return 0                                    # ya expandida
    delta = r_destino - r_ult

    # 1) mover la cola hacia abajo, de abajo arriba. Las combinaciones se
    #    DESHACEN todas primero y se rehacen al final: si se rehace la del
    #    destino antes de copiar, las celdas de llegada ya son `MergedCell` y
    #    `dst.value = …` revienta con «object attribute 'value' is read-only»
    #    (medido con el pie `01!Cuadrante Semanal!A23:J23`).
    rehacer = []
    for m in [str(r) for r in ws.merged_cells.ranges]:
        partes = m.split(':')
        mm = RX_REF.fullmatch(partes[0])
        if mm and int(mm.group(4)) in cola:
            ws.unmerge_cells(m)
            rehacer.append(':'.join(_corre_ref(p, r_ult, delta)
                                    for p in partes))
    for origen in sorted(cola, reverse=True):
        for c in range(1, max_c + 1):
            src = ws.cell(row=origen, column=c)
            dst = ws.cell(row=origen + delta, column=c)
            dst.value = _corre_cola(src.value, r_ult, r_destino, delta)
            dst._style = copy.copy(src._style)
            src.value = None
    for m in rehacer:
        ws.merge_cells(m)

    # 2) replicar la última fila del bloque hacia abajo
    base = ws.cell(row=r_ult, column=numerar).value if numerar else None
    for fila in range(r_ult + 1, r_destino + 1):
        plantilla = fila - 2 if fila - 2 >= 1 else r_ult
        for c in range(1, max_c + 1):
            src = ws.cell(row=r_ult, column=c)
            dst = ws.cell(row=fila, column=c)
            dst._style = copy.copy(ws.cell(row=plantilla, column=c)._style)
            if isinstance(src.value, str) and src.value.startswith('='):
                dst.value = _corre_filas(src.value, 0, fila - r_ult)
                _reg(ws, dst.coordinate, dst.value)
            elif numerar and c == numerar and isinstance(base, int):
                dst.value = base + (fila - r_ult)
            else:
                dst.value = None
    return delta


def _corre_ref(ref, desde, delta):
    m = RX_REF.fullmatch(ref)
    if not m:
        return ref
    d1, col, d2, fila = m.groups()
    fi = int(fila)
    return '{}{}{}{}'.format(d1, col, d2, fi + delta if fi > desde else fi)


def _corre_cola(valor, r_ult, r_destino, delta):
    """Traduce una fórmula de la COLA al expandir el bloque. Dos reglas: una
    referencia POR DEBAJO del bloque se desplaza `delta`; una referencia a la
    ÚLTIMA fila del bloque se ESTIRA hasta la nueva última. Sin la segunda,
    `01!Cuadrante Semanal!I21 = SUM(I6:I20)` seguiría sumando 15 empleados
    después de crecer a 30 y el total del equipo mentiría."""
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        fi = int(fila)
        if fi == r_ult:
            fila = str(r_destino)
        elif fi > r_ult:
            fila = str(fi + delta)
        return '{}{}{}{}'.format(d1, col, d2, fila)

    return RX_REF.sub(_sub, valor)


def _corre_filas(valor, desde, delta):
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        fi = int(fila)
        if fi > desde:
            fila = str(fi + delta)
        return '{}{}{}{}'.format(d1, col, d2, fila)

    return RX_REF.sub(_sub, valor)


def print_setup(ws, header_row=None, landscape=True, fit_to_width=1,
                title_cols=None):
    """A4 con pie de página. El censo cuenta como defecto (`noprint`) toda hoja
    con `paperSize != 9`."""
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = '{0}:{0}'.format(header_row)
    if title_cols:
        ws.print_title_cols = title_cols


def col_instrucciones(ws):
    """En ESTE kit el texto de `Instrucciones` vive en la columna **B** (los
    anchos son `A=3, B=80`), no en la A como en el kit-inventario. Un
    `linea_instrucciones` que escribiera en A dejaría la bio en una columna de 3
    caracteres de ancho, invisible. Se deduce: gana la columna con más cadenas.
    """
    mejor, cuantas = 1, -1
    for c in range(1, min(4, ws.max_column) + 1):
        n = sum(1 for r in range(1, ws.max_row + 1)
                if isinstance(ws.cell(row=r, column=c).value, str))
        if n > cuantas:
            mejor, cuantas = c, n
    return mejor


def linea_instrucciones(ws, texto, rx=None, col=None):
    """Escribe `texto` en Instrucciones: sustituye la línea que case con `rx` o
    la añade al final si no existe. Nunca duplica."""
    if col is None:
        col = col_instrucciones(ws)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return r
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return r
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)
    return destino


# ==========================================================================
# Bloques y centinelas
# ==========================================================================
def _bloques(fname, hoja):
    return [e for e in BLOQUES.get(fname, []) if e[0] == hoja]


def secciones_04(ws):
    """Los 5 tramos del checklist del 04, DETECTADOS (no tabulados).

    Hoy los datos van 7-16 / 20-27 / 31-39 / 43-54 / 58-65, pero grupo_c añade
    tres tareas al primer tramo (Contrat@ al SEPE, copia básica a la RLT y
    modelo 145 — DOM-15) y todo lo de abajo baja tres filas. Una tabla fija
    quedaría desfasada EN LA MISMA PASADA que las añade, y el contador acotado
    por tramo —que es la razón de ser del §4— sumaría los tramos equivocados.

    Devuelve [(fila_titulo, fila_cabecera, r0, r1)]. El título de sección es una
    celda combinada de la columna A con emoji; la cabecera es la fila siguiente
    (la que dice «#»/«Tarea»); los datos van hasta dos filas antes del título
    siguiente (hay una fila en blanco de separación).
    """
    titulos = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        for clave, _color in SECCIONES_04:
            if clave in v.upper():
                titulos.append((r, clave))
                break
    fuera = []
    for i, par in enumerate(titulos):
        r_tit = par[0]
        hdr = r_tit + 1
        cab = ws.cell(row=hdr, column=2).value
        if not (isinstance(cab, str) and 'tarea' in cab.lower()):
            continue
        if i + 1 < len(titulos):
            fin = titulos[i + 1][0] - 2
        else:
            fin = None
            for r in range(hdr + 1, ws.max_row + 1):
                if ws.cell(row=r, column=2).value is None:
                    fin = r - 1
                    break
            if fin is None:
                fin = ws.max_row
        if fin >= hdr + 1:
            fuera.append((r_tit, hdr, hdr + 1, fin))
    return fuera


def _fila_cola(ws, desde):
    """Primera fila >= `desde` que ABRE la cola del bloque: un rótulo de texto
    en la columna A que no sea una fórmula («TOTAL HORAS EQUIPO», «TOTALES»,
    «CARNETS Y CERTIFICADOS A RENOVAR», «COBERTURA: QUIÉN SUSTITUYE A QUIÉN»,
    «NOTAS:», el pie «© 2026 AI Chef Pro»…).

    Se busca SÓLO por debajo de la última fila conocida de la v1.1, nunca desde
    `r0`: dentro del bloque la columna A puede llevar texto perfectamente
    legítimo —`05!Cobertura!A6:A8` son «Mañana / Tarde / Noche» y
    `07!Vencimientos!A7:A21` son fórmulas `=IF(Plantilla!A5…)`—, así que
    buscarla desde arriba cerraría el bloque en su primera fila de datos.
    """
    for f in range(desde, ws.max_row + 1):
        v = ws.cell(row=f, column=1).value
        if isinstance(v, str) and v.strip() and not v.startswith('='):
            return f
    return ws.max_row + 1


def _rangos_filas(ws, fname, hoja):
    """Devuelve [(hdr, r0, r1)] de los bloques de DATOS, sin tragarse la cola.

    `ws.max_row` NO sirve como tope: incluye el «TOTAL HORAS EQUIPO» de
    `01!Cuadrante Semanal!A21` y el pie «© 2026 AI Chef Pro». Y la regla del kit
    hermano —«si `ws.max_row` supera `ULTIMA_V2`, el grupo expandió»— aquí
    MIENTE: `07!Vencimientos` llega hoy a la fila 38 por su segundo bloque y su
    pie, sin que nadie haya expandido nada, así que el primer bloque se estiraba
    de `E7:E21` a `E7:E36` y el semáforo de contratos se comía las diez filas de
    carnets. Medido el 2026-08-24.

    Regla correcta, y además dinámica: el bloque llega hasta la fila anterior a
    su COLA, y sólo crece por encima de `ULTIMA_HOY` mientras las filas tengan
    contenido (que es lo que deja `expandir_filas` al replicar las fórmulas).
    Como la cola se localiza por su rótulo y no por su número de fila, esto
    sigue siendo cierto después de que un grupo añada 250 filas.
    """
    if fname.startswith('04-') and hoja == 'Checklist Onboarding':
        return [(hdr, r0, r1) for _t, hdr, r0, r1 in secciones_04(ws)]
    fuera = []
    for _h, hdr, r0, r1_hoy, r1_v2 in _bloques(fname, hoja):
        tope = _fila_cola(ws, r1_hoy + 1)
        techo = min(r1_v2, tope - 1)
        r1 = min(r1_hoy, techo)
        f = r1 + 1
        while f <= techo and any(ws.cell(row=f, column=c).value is not None
                                 for c in range(1, ws.max_column + 1)):
            r1 = f
            f += 1
        if r1 >= r0:
            fuera.append((hdr, r0, r1))
    return fuera


def _dice(ws, coord, texto):
    """Centinela de CELDA FIJA: `ws!coord` ya contiene `texto`.

    ⚠ NO se puede escribir `if coord in ws`. `Worksheet.__contains__` indexa por
    la TUPLA `(fila, columna)`, no por la coordenada en texto, así que
    `'A3' in ws` es **siempre False** — y con ello se saltaban en silencio los
    tres semáforos de celda fija (`03!Ratio Coste Laboral!B9`,
    `06!Ficha Evaluación!C23`, la fila 38 del calendario) y los cuatro rangos de
    formato de moneda de `RANGOS_FORMATO`: `02!Resumen Mensual!B3` seguía
    enseñando «12» donde tiene que poner «12,00 €». El gate no lo cazaba porque
    un centinela que devuelve False es indistinguible de «el grupo todavía no lo
    ha construido». Medido el 2026-08-24.
    """
    try:
        v = ws[coord].value
    except (ValueError, KeyError, IndexError):
        return False
    return isinstance(v, str) and texto.lower() in v.lower()


def _cabecera_dice(ws, hdr, col, texto):
    """Centinela: la cabecera de ESTE bloque ya dice lo que tiene que decir.

    ⚠ La comprobación de límites NO es defensiva: `ws.cell(row, column)` de
    openpyxl **crea** la celda si no existe, y con ella agranda `ws.max_column`.
    Preguntar por `07!Vencimientos!I6` —una columna que sólo existirá cuando
    grupo_c monte los 4 vencimientos— materializaba las columnas F..I vacías y
    el `print_area` pasaba de `A1:E38` a `A1:I38`: cuatro columnas en blanco
    dentro del área de impresión de una hoja de alertas. Medido el 2026-08-24.
    """
    if texto is None:
        return True
    ci = column_index_from_string(col)
    if ci > ws.max_column or hdr > ws.max_row:
        return False
    v = ws.cell(row=hdr, column=ci).value
    return isinstance(v, str) and texto.lower() in v.lower()


# ==========================================================================
# §1.1 — validación de datos
# ==========================================================================
def _dv_inline(valores, titulo, prompt, error):
    """DV de lista INLINE. Excel limita `formula1` a 255 caracteres contando las
    comillas: los 8 códigos de jornada ocupan 17 y las 4 modalidades de hora
    extra 56, así que caben de sobra. Si algún día no cupieran, ABORTA en vez de
    escribir una lista truncada."""
    formula = '"{}"'.format(','.join(valores) if isinstance(valores, list)
                            else valores)
    if len(formula) > 255:
        raise ValueError('DV inline de {} caracteres (>255): {}'
                         .format(len(formula), titulo))
    return DataValidation(
        type='list', formula1=formula, allow_blank=True,
        showErrorMessage=True, errorTitle=titulo, error=error,
        errorStyle='stop', showInputMessage=True,
        promptTitle='{} · {}'.format(MARCA_DV, titulo), prompt=prompt)


def _celdas_sqref(sqref):
    """Conjunto de (columna, fila) que cubre un sqref, acotado para no explotar
    con un rango entero de columna."""
    fuera = set()
    for token in str(sqref).replace(',', ' ').split():
        partes = token.split(':')
        m0 = RX_REF.fullmatch(partes[0])
        if not m0:
            continue
        c0 = column_index_from_string(m0.group(2))
        f0 = int(m0.group(4))
        if len(partes) == 1:
            fuera.add((c0, f0))
            continue
        m1 = RX_REF.fullmatch(partes[1])
        if not m1:
            fuera.add((c0, f0))
            continue
        c1 = column_index_from_string(m1.group(2))
        f1 = int(m1.group(4))
        if (c1 - c0 + 1) * (f1 - f0 + 1) > 60000:
            continue
        for c in range(c0, c1 + 1):
            for f in range(f0, f1 + 1):
                fuera.add((c, f))
    return fuera


def _limpiar_dv(ws, objetivo=None):
    """Quita las DV del motor (las reconoce por `promptTitle`) y, si se le pasa
    un `objetivo` (conjunto de celdas), también las HEREDADAS de la v1.1 que
    pisen ese rango.

    Lo segundo no es opcional: `01!Cuadrante Semanal` trae una DV
    `"M,T,N,P,L"` sobre `B6:H20` (medida en el fichero real). Si el motor
    añadiera la suya —`M,T,N,P,D,L,V,B`— sin quitar aquélla, la celda tendría
    DOS validaciones y Excel aplica la primera que encuentra: el desplegable
    seguiría sin ofrecer el turno doble, con la regla nueva escrita y todo.
    """
    vivos, quitadas = [], 0
    for dv in ws.data_validations.dataValidation:
        pt = getattr(dv, 'promptTitle', None) or ''
        if pt.startswith(MARCA_DV):
            quitadas += 1
            continue
        if objetivo and (_celdas_sqref(dv.sqref) & objetivo):
            quitadas += 1
            continue
        vivos.append(dv)
    ws.data_validations.dataValidation = vivos
    return quitadas


def aplicar_dv(wb, fname, informe, pendientes):
    """§1.1 — DV de códigos (leyenda única) y DV de listas."""
    puestas = 0
    objetivos = {}     # hoja → (conjunto de celdas, [(ref, dv)])

    for hoja, c_desde, c_hasta, texto, familia in DV_CODIGOS.get(fname, []):
        if hoja not in wb.sheetnames:
            pendientes.append('{}:{}: la hoja no existe todavía'
                              .format(fname, hoja))
            continue
        ws = wb[hoja]
        rangos = _rangos_filas(ws, fname, hoja)
        if not rangos:
            pendientes.append('{}:{}: sin bloque de datos'.format(fname, hoja))
            continue
        hdr, r0, r1 = rangos[0]
        if not _cabecera_dice(ws, hdr, c_desde, texto):
            pendientes.append(
                '{}:{}:{}{}: la cabecera no dice «{}» todavía — DV de códigos '
                'aplazada al grupo'.format(fname, hoja, c_desde, hdr, texto))
            continue
        ultima = c_hasta
        if ultima is None:
            # la última columna del calendario es el TOTAL, no un día
            ultima = get_column_letter(max(2, ws.max_column - 1))
        if familia == 'jornada':
            valores, titulo = DV_JORNADA, 'Código de turno no válido'
            prompt = LEYENDA_JORNADA + '.'
            error = ('Usa uno de los 8 códigos del kit. Son los mismos en el '
                     'cuadrante semanal y en el mensual, y de ellos salen las '
                     'horas y las alertas: una letra inventada cuenta 0 h sin '
                     'avisar.')
        else:
            valores, titulo = DV_AUSENCIA, 'Código de ausencia no válido'
            prompt = LEYENDA_AUSENCIA + '.'
            error = ('Usa V, B, F o PE. El permiso es PE: la P está ocupada '
                     'por el turno Partido en el cuadrante del 01.')
        dv = _dv_inline(valores, titulo, prompt, error)
        ref = '{a}{f0}:{b}{f1}'.format(a=c_desde, b=ultima, f0=r0, f1=r1)
        objetivos.setdefault(hoja, [set(), []])
        objetivos[hoja][0] |= _celdas_sqref(ref)
        objetivos[hoja][1].append((ref, dv, '{} códigos de {}'
                                   .format(len(valores.split(',')), familia)))

    for entrada in DV_LISTA.get(fname, []):
        hoja, col, texto, valores, titulo, prompt = entrada
        if hoja not in wb.sheetnames:
            pendientes.append('{}:{}: la hoja no existe todavía'
                              .format(fname, hoja))
            continue
        ws = wb[hoja]
        colocada = False
        for hdr, r0, r1 in _rangos_filas(ws, fname, hoja):
            if not _cabecera_dice(ws, hdr, col, texto):
                continue
            dv = _dv_inline(valores, titulo, prompt,
                            'Elige un valor de la lista: las fórmulas de '
                            'agregación buscan este texto exacto.')
            ref = '{c}{a}:{c}{b}'.format(c=col, a=r0, b=r1)
            objetivos.setdefault(hoja, [set(), []])
            objetivos[hoja][0] |= _celdas_sqref(ref)
            objetivos[hoja][1].append((ref, dv, '{} valores'
                                       .format(len(valores))))
            colocada = True
        if not colocada:
            pendientes.append(
                '{}:{}!{}: ningún bloque tiene «{}» en la cabecera — DV '
                'aplazada al grupo'.format(fname, hoja, col, texto))

    for hoja, par in objetivos.items():
        ws = wb[hoja]
        quitadas = _limpiar_dv(ws, par[0])
        for ref, dv, detalle in par[1]:
            ws.add_data_validation(dv)
            dv.add(ref)
            puestas += 1
            informe.append('{}:{}!{}: DV inline ({}){}'
                           .format(fname, hoja, ref, detalle,
                                   '' if not quitadas else
                                   ' — {} DV heredadas retiradas del rango'
                                   .format(quitadas)))
            quitadas = 0
    return puestas


# ==========================================================================
# §1.2 — formato condicional REAL
# ==========================================================================
def _dxf(color):
    bg, fg = {'verde': (CF_VERDE_BG, CF_VERDE_FG),
              'ambar': (CF_AMBAR_BG, CF_AMBAR_FG),
              'rojo': (CF_ROJO_BG, CF_ROJO_FG),
              'azul': (CF_AZUL_BG, CF_AZUL_FG),
              'naranja': (CF_NARANJA_BG, CF_NARANJA_FG),
              'gris': (CF_GRIS_BG, CF_GRIS_FG)}[color]
    return DifferentialStyle(font=Font(color=fg, bold=True),
                             fill=PatternFill(start_color=bg, end_color=bg,
                                              fill_type='solid'))


def _norm_ref(ref):
    """`B9:B9` → `B9`. openpyxl NORMALIZA el sqref de un rango de una sola celda
    al guardarlo, así que la 2.ª pasada leía `B9` y lo comparaba contra el
    `B9:B9` que el motor iba a escribir: `_limpiar_cf` no reconocía sus propias
    reglas, no las borraba, y las 3 de `03!Ratio Coste Laboral!B9` y las 5 de
    `06!Ficha Evaluación!C23` se DUPLICABAN en cada pasada (6 y 10 en la
    segunda). Medido el 2026-08-24: era la única diferencia de idempotencia que
    quedaba, y habría ido creciendo en cada ejecución hasta que Excel se
    quejara del número de reglas."""
    partes = ref.split(':')
    if len(partes) == 2 and partes[0] == partes[1]:
        return partes[0]
    return ref


def _limpiar_cf(ws, refs):
    """Borra las reglas de los rangos que gobierna el motor y deja intactas las
    de los grupos. `refs` = conjunto de sqref en texto."""
    normalizadas = set(_norm_ref(r) for r in refs)
    nueva = ConditionalFormattingList()
    for cf in ws.conditional_formatting:
        if _norm_ref(str(cf.sqref)) in normalizadas:
            continue
        for regla in cf.rules:
            nueva.add(str(cf.sqref), regla)
    ws.conditional_formatting = nueva


def semaforo(ws, rango, vocabulario, stop=True):
    """§1.2 — pinta un rango de TEXTO según las palabras del vocabulario.

    Hoy las 22 hojas del kit tienen UNA sola regla de formato condicional
    (`04-onboarding-nuevo-empleado.xlsx:Checklist Onboarding:A7:G65`): el
    «semáforo» que prometen las Instrucciones de los otros ocho ficheros es el
    emoji dentro de la cadena. Esto lo hace real, y se sigue leyendo en una
    impresión en blanco y negro porque el emoji sigue ahí.
    """
    primera = RX_REF.fullmatch(rango.split(':')[0])
    col, fila = primera.group(2), primera.group(4)
    for texto, color in vocabulario:
        formula = 'NOT(ISERROR(SEARCH("{t}",{c}{f})))'.format(
            t=texto, c=col, f=fila)
        ws.conditional_formatting.add(rango, Rule(
            type='containsText', operator='containsText', text=texto,
            dxf=_dxf(color), stopIfTrue=stop, formula=[formula]))
    return len(vocabulario)


def _voc_codigos():
    """Vocabulario del calendario del 05: cada código con su color de §1.2.
    `PE` va PRIMERO y con `stopIfTrue`: si se evaluara «F» antes, un permiso no
    se pintaría de naranja porque `SEARCH` no distingue mayúsculas y `PE` no
    contiene `F`… pero sí contiene `E`, y un vocabulario mal ordenado en este
    punto es exactamente el error que costó el `VOC_COTIZACION` del kit
    hermano («VENCIDA» no contiene «vence»)."""
    mapa = {'BBDEFB': 'azul', 'FFC7CE': 'rojo', 'C8E6C9': 'verde',
            'FFE0B2': 'naranja'}
    orden = sorted(CODIGOS_AUSENCIA, key=lambda c: -len(c[0]))
    return [(c[0], mapa[c[2]]) for c in orden]


def aplicar_cf(wb, fname, informe, pendientes):
    """§1.2 — semáforos por texto, semáforos de celda y reglas de expresión."""
    puestas = 0
    porhoja = {}

    for hoja, col, texto, voc in CF_COLUMNA.get(fname, []):
        porhoja.setdefault(hoja, []).append(('col', col, texto, voc, None,
                                             None))
    for hoja, ref, cc, ct, voc in CF_CELDA.get(fname, []):
        porhoja.setdefault(hoja, []).append(('celda', ref, cc, ct, voc, None))
    for hoja, ref, cc, ct, expr, color, stop in CF_EXPRESION.get(fname, []):
        porhoja.setdefault(hoja, []).append(('expr', ref, cc, ct, expr,
                                             (color, stop)))

    # el calendario del 05 lleva además el semáforo de CÓDIGOS sobre la rejilla
    if fname.startswith('05-'):
        porhoja.setdefault('Calendario Anual', []).append(
            ('codigos', 'B', 'Semana', _voc_codigos(), None, None))

    for hoja, objetivos in porhoja.items():
        if hoja not in wb.sheetnames:
            pendientes.append('{}:{}: la hoja no existe todavía'
                              .format(fname, hoja))
            continue
        ws = wb[hoja]
        rangos = _rangos_filas(ws, fname, hoja)
        refs, listos = set(), []
        for tipo, a, b, c, d, extra in objetivos:
            if tipo == 'col':
                col, texto, voc = a, b, c
                encontrada = False
                for hdr, r0, r1 in rangos:
                    if not _cabecera_dice(ws, hdr, col, texto):
                        continue
                    ref = '{c}{a}:{c}{b}'.format(c=col, a=r0, b=r1)
                    refs.add(ref)
                    listos.append(('texto', ref, voc, None))
                    encontrada = True
                if not encontrada:
                    pendientes.append(
                        '{}:{}!{}: ningún bloque dice «{}» en la cabecera — '
                        'formato condicional aplazado al grupo'
                        .format(fname, hoja, col, texto))
            elif tipo == 'codigos':
                # el color de los códigos se aplica SOBRE LA REJILLA QUE HAYA:
                # hoy es la de 12 meses y mañana la de 53 semanas, y en las dos
                # el valor de la celda es la misma letra. Condicionarlo a que
                # grupo_c hubiera reconstruido el calendario habría dejado el
                # `--solo motor` sin la única mejora visible del 05 —hoy la
                # hoja tiene CERO reglas de formato condicional pese a que sus
                # Instrucciones prometen «marca con código de color»
                # (`05!Instrucciones!B5`, COM-10)—.
                col, texto, voc = a, b, c
                if not rangos:
                    pendientes.append('{}:{}: sin bloque de datos'
                                      .format(fname, hoja))
                    continue
                hdr, r0, r1 = rangos[0]
                ultima = get_column_letter(max(2, ws.max_column - 1))
                ref = '{a}{f0}:{b}{f1}'.format(a=col, b=ultima, f0=r0, f1=r1)
                refs.add(ref)
                listos.append(('texto', ref, voc, None))
            elif tipo == 'celda':
                ref, cc, ct, voc = a, b, c, d
                if not _dice(ws, cc, ct):
                    pendientes.append(
                        '{}:{}!{}: no dice «{}» todavía — formato condicional '
                        'aplazado al grupo'.format(fname, hoja, cc, ct))
                    continue
                refs.add(ref)
                listos.append(('texto', ref, voc, None))
            else:
                ref, cc, ct, expr = a, b, c, d
                if not _dice(ws, cc, ct):
                    pendientes.append(
                        '{}:{}!{}: no dice «{}» todavía — banda por mes '
                        'aplazada al grupo'.format(fname, hoja, cc, ct))
                    continue
                refs.add(ref)
                listos.append(('expr', ref, expr, extra))
        # RT-04 · las reglas de EXPRESIÓN van SIEMPRE las últimas.
        #
        # openpyxl numera la prioridad por orden de inserción y Excel resuelve
        # un conflicto a favor de la regla de MENOR número. La banda gris por
        # mes del calendario del 05 y las cuatro reglas de color de código
        # escriben la MISMA propiedad —el relleno—, así que insertar la banda
        # antes le daba precedencia: medido en la copia dry-run, prioridades
        # 3 (banda, `stopIfTrue=False`) frente a 4-7 (PE, V, B, F). En los
        # meses IMPARES —enero, marzo, mayo, julio, septiembre, noviembre— la
        # banda ganaba y una V, una B, una F o un PE se pintaban GRISES.
        # Julio y septiembre son justamente las semanas 27-35 que el propio
        # fichero precarga como temporada alta: el semáforo se anulaba medio
        # año en la hoja donde más se ve.
        #
        # El orden viene del orden de `objetivos` (CF_COLUMNA, CF_CELDA,
        # CF_EXPRESION y por último los códigos), así que se reordena aquí en
        # vez de confiar en cómo se construyó la lista. `sort` es estable: el
        # resto conserva su orden.
        listos.sort(key=lambda x: x[0] == 'expr')
        _limpiar_cf(ws, refs)
        for tipo, ref, dato, extra in listos:
            if tipo == 'texto':
                n = semaforo(ws, ref, dato)
            else:
                color, stop = extra
                f0 = RX_REF.fullmatch(ref.split(':')[0]).group(4)
                ws.conditional_formatting.add(ref, Rule(
                    type='expression', dxf=_dxf(color), stopIfTrue=stop,
                    formula=[dato.format(f=f0).lstrip('=')]))
                n = 1
            puestas += n
            informe.append('{}:{}!{}: {} regla(s) de formato condicional'
                           .format(fname, hoja, ref, n))
    return puestas


# ==========================================================================
# §1.1 — leyenda en la fila 3 de cada rejilla y en las Instrucciones
# ==========================================================================
def escribir_leyenda(wb, fname, informe):
    """§1.1 — la fila 3 de cada rejilla enseña TODOS los códigos con su color.

    Hoy `01!Cuadrante Semanal!B3:F3` enseña 5 de 8 (faltan D, V y B) y
    `05!Calendario Anual!E3` dice «P=Permiso», que choca de frente con el
    «P (partido)» de `01!Instrucciones!B6`. Un código con dos significados en el
    mismo kit es un cuadrante mal leído.
    """
    n = 0
    for hoja, fila, familia in LEYENDA_REJILLA.get(fname, []):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        codigos = (CODIGOS_JORNADA if familia == 'jornada'
                   else CODIGOS_AUSENCIA)
        # limpiar la leyenda vieja de la fila entera antes de reescribir: la del
        # 05 ocupa B3:E3 y la nueva B3:E3 también, pero la del 01 pasa de 5 a 8
        # celdas y dejar restos daría dos leyendas contradictorias.
        for c in range(2, ws.max_column + 1):
            cel = ws.cell(row=fila, column=c)
            if isinstance(cel.value, str) and '=' in cel.value:
                cel.value = None
                cel.fill = PatternFill()
        for i, cod in enumerate(codigos):
            cel = ws.cell(row=fila, column=2 + i)
            cel.value = '{}={}'.format(cod[0], cod[1])
            cel.fill = PatternFill('solid', fgColor=cod[-1])
            cel.font = Font(bold=True, size=9)
            cel.alignment = Alignment(horizontal='center')
            n += 1
        informe.append('{}:{}!B{}:{}{}: leyenda única de {} códigos (§1.1)'
                       .format(fname, hoja, fila,
                               get_column_letter(1 + len(codigos)), fila,
                               len(codigos)))
    if 'Instrucciones' in wb.sheetnames:
        ws = wb['Instrucciones']
        for f, rx, texto in LINEAS_LEYENDA:
            if f != fname:
                continue
            fila = linea_instrucciones(ws, texto, rx)
            informe.append('{}:Instrucciones!{}{}: leyenda única (§1.1)'
                           .format(fname, get_column_letter(
                               col_instrucciones(ws)), fila))
            n += 1
    return n


# ==========================================================================
# 04 — colores de sección y desactivación de la colisión con el verde
# ==========================================================================
def colores_seccion_04(wb, fname, informe):
    """§1.2 — pinta las 5 cabeceras de sección del 04 con los colores que sus
    propias Instrucciones anuncian, y saca el `E8F5E9` de la columna
    «Categoría» (ver `SECCIONES_04`)."""
    if not fname.startswith('04-') or 'Checklist Onboarding' not in \
            wb.sheetnames:
        return 0
    ws = wb['Checklist Onboarding']
    n = 0
    for r_tit, hdr, r0, r1 in secciones_04(ws):
        titulo = (ws.cell(row=r_tit, column=1).value or '').upper()
        color = None
        for clave, col in SECCIONES_04:
            if clave in titulo:
                color = col
                break
        if not color:
            continue
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r_tit, column=c)
            if _relleno(cel) != color:
                cel.fill = PatternFill('solid', fgColor=color)
                n += 1
        # la columna «Categoría» pierde el verde de edición
        for fila in range(r0, r1 + 1):
            cel = ws.cell(row=fila, column=3)
            if es_verde(cel):
                cel.fill = PatternFill('solid', fgColor=color)
                cel.protection = Protection(locked=True)
                n += 1
    if n:
        informe.append('{}:Checklist Onboarding: {} celdas de sección '
                       'recoloreadas — el «EQUIPAMIENTO» pintado de E8F5E9 '
                       'colisionaba con el verde de EDICIÓN de la familia '
                       '(§1.2)'.format(fname, n))
    return n


# ==========================================================================
# §1.6 — formatos
# ==========================================================================
#: RT-20 · cabeceras cuyo contenido es TEXTO aunque su palabra caiga en una de
#: las familias de arriba. El caso medido: `07!'Plantilla'!G` se llama
#: «Jornada» —que el mapa envía a `0.00` por la familia de las horas— y su DV
#: sólo admite las cadenas «Completa» y «Parcial»: un formato numérico sobre
#: una columna donde no se puede escribir un número. Va con coincidencia
#: EXACTA, no por subcadena: «Jornada contratada (h/semana)» sí es numérica.
CABECERAS_DE_TEXTO = frozenset([
    # 07!'Plantilla'!G · DV de «Completa,Parcial»
    'jornada',
    # 01!'Cuadrante Semanal'!K:P y 01!'Cuadrante Mensual'!J · las alertas
    # devuelven «⛔ …» o cadena vacía, y su cabecera lleva la palabra
    # «jornada», que el mapa envía a 0.00.
    'descanso entre jornadas', 'descanso semanal', 'jornada semanal',
    'jornada diaria', 'menor de edad (s/n)', 'alerta menores',
    'alerta semanal', 'alerta del cómputo', 'alerta del computo',
    # 02!'Resumen Mensual'!F · veredicto frente al tope anual (RT-16)
    'estado frente al límite anual', 'estado frente al limite anual',
    # BONUS-02!'Ratios por Tipo'!H · «28-33%», texto
    'ratio coste laboral',
    # genéricas
    'estado', 'aviso', 'alerta', 'conforme', 'veredicto', 'plazo',
    'tendencia', 'nivel', 'semáforo', 'semaforo'])


def _formato_de_cabecera(texto):
    if not isinstance(texto, str):
        return None
    t = texto.lower().strip()
    # RT-20 · «Cubiertos/cocinero y servicio (rango)» contiene «cubiertos» y el
    # mapa lo mandaba a entero, pero su contenido es «22-28» o «—».
    if t in CABECERAS_DE_TEXTO or t.endswith('(rango)'):
        return 'General'
    for fmt, claves in FORMATO_POR_CABECERA:
        for k in claves:
            if k in t:
                return fmt
    return None


def aplicar_formatos(wb, fname, informe):
    """§1.6 — hora, fecha, moneda, porcentaje y entero, deducidos de la cabecera.

    El `hh:mm` de `02!C:D` no es cosmética: hoy la columna va en `General` y
    teclear «9» donde se espera «9:00» graba el número 9, que a `(D5-C5)*24` le
    sale **192 h** de jornada sin un solo aviso (TEC-25).
    """
    n = 0
    hojas = set(e[0] for e in BLOQUES.get(fname, []))
    if fname.startswith('04-'):
        hojas.add('Checklist Onboarding')
    for hoja in hojas:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        for hdr, r0, r1 in _rangos_filas(ws, fname, hoja):
            for c in range(1, ws.max_column + 1):
                fmt = _formato_de_cabecera(ws.cell(row=hdr, column=c).value)
                if not fmt:
                    continue
                for fila in range(r0, r1 + 1):
                    cel = ws.cell(row=fila, column=c)
                    if cel.number_format != fmt:
                        cel.number_format = fmt
                        n += 1
    for hoja, ref, fmt, cc, ct in RANGOS_FORMATO.get(fname, []):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        if not _dice(ws, cc, ct):
            continue
        filas = ws[ref]
        if not isinstance(filas, tuple):
            filas = ((filas,),)
        for fila in filas:
            for cel in (fila if isinstance(fila, tuple) else (fila,)):
                if cel.number_format != fmt:
                    cel.number_format = fmt
                    n += 1
    if n:
        informe.append('{}: {} celdas reformateadas (§1.6)'.format(fname, n))
    return n


# ==========================================================================
# Verde de edición
# ==========================================================================
def aplicar_verde(wb, fname, informe):
    """Convención de familia: verde `E8F5E9` = lo escribe el cliente.

    Regla: dentro del bloque de datos, una columna es editable si tiene cabecera
    y NINGUNA de sus celdas contiene fórmula. Las cabeceras de
    `NO_VERDE_CABECERAS` quedan fuera (son contenido impreso: las 50 tareas del
    onboarding, las 10 competencias, la taxonomía de la tabla de ratios) y las
    hojas de `SIN_VERDE_AUTO` las marcan los grupos a mano.

    Y quita el verde de una columna que haya pasado a CALCULADA: cuando grupo_a
    convierta `02!Resumen Mensual!B` en el `SUMIF` que hoy es transcripción
    manual, dejarla verde le diría al cliente que puede sobrescribir la fórmula.
    """
    n, limpiadas = 0, 0
    hojas = set(e[0] for e in BLOQUES.get(fname, []))
    if fname.startswith('04-'):
        hojas.add('Checklist Onboarding')
    for hoja in hojas:
        if hoja not in wb.sheetnames or hoja in SIN_VERDE_AUTO:
            continue
        ws = wb[hoja]
        for hdr, r0, r1 in _rangos_filas(ws, fname, hoja):
            for c in range(1, ws.max_column + 1):
                cab = ws.cell(row=hdr, column=c).value
                if not isinstance(cab, str) or not cab.strip():
                    continue
                if cab.strip().lower() in NO_VERDE_CABECERAS:
                    continue
                celdas = [ws.cell(row=f, column=c) for f in range(r0, r1 + 1)]
                calculada = any(isinstance(x.value, str)
                                and x.value.startswith('=') for x in celdas)
                if c == 1 and all(x.value is not None for x in celdas):
                    calculada = True       # columna de etiquetas impresa
                for cel in celdas:
                    if calculada:
                        if es_verde(cel):
                            cel.fill = copy.copy(
                                ws.cell(row=cel.row, column=1).fill)
                            limpiadas += 1
                    elif not es_verde(cel):
                        cel.fill = PatternFill('solid', fgColor=VERDE)
                        n += 1
    if n or limpiadas:
        informe.append('{}: {} celdas marcadas como editables (verde {}), '
                       '{} desmarcadas por pasar a calculadas'
                       .format(fname, n, VERDE, limpiadas))
    return n


# ==========================================================================
# §1.6 — protección sin contraseña
# ==========================================================================
def celdas_no_ancla(ws):
    """RT-23 — coordenadas NO ancla de cada región combinada.

    Medido el 2026-08-24: la copia dry-run traía **55 celdas desbloqueadas sin
    verde** (23 en `06!'Ficha Evaluación'` —`D4:D8`, `C27:D29`, `C32:D34`,
    `C37:D39`—, otras 23 en `06!'Ficha (ejemplo relleno)'` y 9 en
    `BONUS-01!'Briefing'!B61:D63`). Todas son celdas NO ancla de un rango
    combinado, y **openpyxl no puede arreglarlo**: probado en este mismo
    fichero, asignar `fill` y `Protection(locked=True)` a una `MergedCell`
    funciona en memoria y se PIERDE al guardar (openpyxl 3.1.3 escribe el
    `s=` pero al releer devuelve `locked=False` y `fill=None`).

    En Excel el efecto es nulo: una región combinada se comporta según su
    celda ANCLA —es la que se selecciona al hacer clic y la que gobierna el
    bloqueo—, y se pinta con su estilo. Así que la convención de familia se
    precisa en vez de romperse: «en una hoja protegida, sólo las celdas verdes
    (o las no ancla de una región cuya ancla es verde) están desbloqueadas».
    El gate de `main.py` usa esta lista para excluirlas del recuento y seguir
    suspendiendo por cualquier OTRA celda editable sin marcar.
    """
    fuera = set()
    for rango in ws.merged_cells.ranges:
        for fila in range(rango.min_row, rango.max_row + 1):
            for col in range(rango.min_col, rango.max_col + 1):
                if fila == rango.min_row and col == rango.min_col:
                    continue
                fuera.add('{}{}'.format(get_column_letter(col), fila))
    return fuera


def proteger(ws, informe):
    """Protección SIN contraseña: se desbloquean SÓLO las celdas verdes.

    Si la hoja no tiene ni una celda verde NO se protege: dejarla bloqueada
    entera es peor que dejarla abierta — el cliente no podría escribir nada y
    creería que el fichero está roto.
    """
    verdes = 0
    for row in ws.iter_rows():
        for c in row:
            if es_verde(c):
                c.protection = Protection(locked=False)
                verdes += 1
            else:
                c.protection = Protection(locked=True)
    if not verdes and ws.title not in PROTEGER_SIN_VERDE:
        ws.protection.sheet = False
        return 0
    ws.protection.sheet = True
    # SIN contraseña. Ojo: NO tocar `password`. `= None` revienta openpyxl y
    # `= ''` escribe el hash de la cadena vacía → Excel pediría contraseña al
    # desproteger, justo lo contrario de lo que dice la nota de Instrucciones.
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.autoFilter = False
    ws.protection.sort = False
    informe.append('{}: protegida sin contraseña ({} celdas verdes editables)'
                   .format(ws.title, verdes))
    return verdes


# ==========================================================================
# §1.7 — bio anclada, línea de versión y nota de desprotección
# ==========================================================================
def bio_y_version(wb, informe):
    """§1.7 — deja SIEMPRE tres líneas seguidas al pie de `Instrucciones`: nota
    de desprotección, bio anclada y línea de versión.

    Es una INSERCIÓN: los 9 ficheros terminan hoy en «Versión 1.1 · agosto 2026
    · aichef.pro/kit-gestion-personal · info@aichef.pro» y NO llevan bio (por
    eso `postprocess-transversal.py`, que sólo SUSTITUYE patrones conocidos,
    nunca se la puso; y tampoco casa con su `RX_BIO_VIEJA`, porque el literal
    dice «desde los 17 años», que no está en ese patrón).

    Idempotencia: el ancla es la NOTA DE DESPROTECCIÓN — si ya está, el bloque
    se reescribe en su sitio. Anclar en la línea de versión desplazaría el
    bloque dos filas en cada pasada.
    """
    if 'Instrucciones' not in wb.sheetnames:
        return 0
    ws = wb['Instrucciones']
    col = col_instrucciones(ws)
    fila_nota = fila_bio = fila_ver = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str):
            continue
        if v == NOTA_DESPROTEGER:
            fila_nota = r
        elif v == BIO:
            fila_bio = r
        elif RX_VERSION.match(v):
            fila_ver = r

    if fila_nota:
        inicio = fila_nota
    elif fila_bio:
        inicio = fila_bio - 1
    elif fila_ver:
        inicio = fila_ver
    else:
        inicio = ws.max_row + 2

    estilo = None
    if fila_ver:
        estilo = ws.cell(row=fila_ver, column=col)._style

    for i, texto in enumerate((NOTA_DESPROTEGER, BIO, VERSION_LINE)):
        cel = ws.cell(row=inicio + i, column=col)
        cel.value = texto
        if estilo is not None:
            cel._style = copy.copy(estilo)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
    # una línea de versión perdida en otra fila sería un duplicado
    for r in range(1, ws.max_row + 1):
        if inicio <= r <= inicio + 2:
            continue
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and (RX_VERSION.match(v) or v == BIO
                                   or v == NOTA_DESPROTEGER):
            ws.cell(row=r, column=col).value = None
    informe.append('Instrucciones!{}{}:{}{}: nota de desprotección + bio '
                   'anclada + versión {}'
                   .format(get_column_letter(col), inicio,
                           get_column_letter(col), inicio + 2, VERSION))
    return 3


# ==========================================================================
# §1.7 — metadata OOXML
# ==========================================================================
def set_metadata(wb, fname, informe):
    """Propiedades que ve el cliente en Archivo → Información. `subject` pasa a
    `… · v2.0`; el resto sigue la convención de `postprocess-transversal.py:350`
    (`creator='AI Chef Pro'`, no 'openpyxl').

    De paso, el `title` recupera la tilde: los 9 ficheros llevan hoy
    «Kit Gestion de Personal y Turnos» en las propiedades mientras su propio
    contenido escribe «Gestión» con tilde (DOM-26).
    """
    p = wb.properties
    titulo = p.title if isinstance(p.title, str) else ''
    if not titulo.strip():
        base = re.sub(r'^(BONUS-)?\d+[a-z]?-', '', fname[:-5])
        titulo = base.replace('-', ' ').capitalize()
    if not titulo.endswith(CORTO):
        titulo = '{} · {}'.format(titulo.split(' · ')[0]
                                  if ' · ' in titulo else titulo, CORTO)
        # los títulos actuales son «NN · Nombre largo · Kit …»: hay que
        # conservar el nombre largo, no sólo el número.
        partes = (p.title or '').split(' · ')
        if len(partes) >= 2:
            titulo = '{} · {}'.format(' · '.join(partes[:-1]), CORTO)
    quiero = dict(
        creator='AI Chef Pro',
        lastModifiedBy='AI Chef Pro',
        title=titulo,
        subject='{} · v{}'.format(CORTO, VERSION),
        keywords='{}, AI Chef Pro'.format(PID.replace('-', ' ')),
        description='aichef.pro/{}'.format(PID),
        category='AI Chef Pro · Productos digitales',
    )
    n = 0
    for campo, valor in quiero.items():
        if getattr(p, campo) != valor:
            setattr(p, campo, valor)
            n += 1
    if n:
        informe.append('{}: {} propiedades OOXML actualizadas (subject → v{})'
                       .format(fname, n, VERSION))
    return n


# ==========================================================================
# Contadores y gates auxiliares
# ==========================================================================
RX_ENTRECOMILLADO = re.compile(r"['‘’\"“”]([^'‘’\"“”]{3,40})['‘’\"“”]")


def _cabeceras(wb):
    """Todos los textos que son ENCABEZADO de columna en alguna hoja. Sirven
    para no confundir «'Empleado'» o «'Alertas'» —que son columnas— con una
    pestaña inexistente."""
    fuera = set()
    for ws in wb.worksheets:
        for fila in range(1, min(12, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=fila, column=c).value
                if isinstance(v, str) and 1 <= len(v) <= 40:
                    fuera.add(v.strip().lower())
    return fuera


def pestanas_citadas(wb):
    """Gate: toda pestaña entrecomillada en `Instrucciones` tiene que existir.

    En este kit falla al menos dos veces, medidas: `03!Instrucciones!B7` manda a
    la hoja «Ratio Coste Laboral» —ésa sí existe— pero `05!Instrucciones!B7`
    cita «Solicitudes» y `06!Instrucciones!B8` cita «Histórico», que existen;
    los grupos van a prometer «Turnos» (01), «Saldo Vacaciones» (05) y «Ficha
    (ejemplo relleno)» (06), que hoy NO existen. El gate es el inventario de
    esas promesas. Devuelve [(fila, nombre_citado)].
    """
    if 'Instrucciones' not in wb.sheetnames:
        return []
    ws = wb['Instrucciones']
    col = col_instrucciones(ws)
    reales = set(h.strip().lower() for h in wb.sheetnames)
    cabeceras = _cabeceras(wb)
    fuera = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str):
            continue
        for m in RX_ENTRECOMILLADO.finditer(v):
            nombre = m.group(1).strip()
            if not nombre or nombre.lower() in reales:
                continue
            if nombre.lower() in cabeceras:
                continue          # es una COLUMNA, no una pestaña
            fuera.append((r, nombre))
    return fuera


def citas_obsoletas(wb):
    """Gate auxiliar §1.7 — afirmaciones que la v1.1 imprime y son falsas.

    NO se corrigen en el motor (el texto de cada fichero es trabajo de su
    grupo): se CUENTAN, para que una pasada `--solo motor` diga cuántas siguen
    vivas en vez de dar un verde que sólo mide lo que el motor toca.
    """
    fuera = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str) or len(c.value) < 4:
                    continue
                for rx, motivo in CITAS_OBSOLETAS:
                    if rx.search(c.value):
                        fuera.append('{}!{}: {} — «{}»'
                                     .format(ws.title, c.coordinate, motivo,
                                             c.value[:80]))
                        break
    return fuera


def hojas_esqueleto(wb, fname):
    """Contador: hojas cuyo bloque de datos está COMPLETAMENTE en blanco. En
    este kit son casi todas —es un kit de plantillas vacías, no de ejemplos— y
    por eso es INFORMATIVO: sirve para que los grupos sepan dónde sembrar."""
    fuera = []
    for hoja in sorted(set(e[0] for e in BLOQUES.get(fname, []))):
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        rangos = _rangos_filas(ws, fname, hoja)
        if not rangos:
            fuera.append(hoja)
            continue
        lleno = False
        for hdr, r0, r1 in rangos:
            for fila in range(r0, r1 + 1):
                for c in range(2, ws.max_column + 1):
                    v = ws.cell(row=fila, column=c).value
                    if v is None:
                        continue
                    if isinstance(v, str) and v.startswith('='):
                        continue
                    lleno = True
                    break
                if lleno:
                    break
            if lleno:
                break
        if not lleno:
            fuera.append(hoja)
    return fuera


def _columnas_de_sqref(sqref):
    """Índices de columna que cubre un `sqref` de validación de datos."""
    fuera = set()
    for trozo in str(sqref).split():
        for parte in trozo.split(','):
            if not parte:
                continue
            extremos = parte.split(':')
            try:
                c0 = column_index_from_string(
                    re.match(r'\$?([A-Z]+)', extremos[0]).group(1))
                c1 = column_index_from_string(
                    re.match(r'\$?([A-Z]+)', extremos[-1]).group(1))
            except (AttributeError, ValueError):
                continue
            fuera |= set(range(min(c0, c1), max(c0, c1) + 1))
    return fuera


def leyenda_coherente(wb, fname):
    """Gate §1.1 — ninguna DV de la REJILLA puede ofrecer una letra que la
    leyenda no explique. Es lo que impide que vuelva el `P` con dos
    significados (DOM-27).

    Se acota a las hojas que llevan códigos de turno o de ausencia. Barrer el
    libro entero daba tres falsos positivos en
    `04-onboarding-nuevo-empleado.xlsx:Checklist Onboarding:F7`, cuya DV
    `"✓,✗,—"` también es de valores de un carácter y no tiene nada que ver con
    la leyenda: un gate que suspende por lo que no mide deja de leerse.
    """
    fallos = []
    validas = set(c[0] for c in CODIGOS_JORNADA) | \
        set(c[0] for c in CODIGOS_AUSENCIA)
    hojas = set(e[0] for e in DV_CODIGOS.get(fname, []))
    hojas |= set(e[0] for e in LEYENDA_REJILLA.get(fname, []))
    #: RD-16 · y se acota además por COLUMNA. Antes bastaba con que una DV
    #: viviera en la misma HOJA de la rejilla para que el gate la juzgara con
    #: la leyenda de códigos: la casilla nueva «Menor de edad (S/N)» del 01,
    #: que está en la O y no tiene nada que ver con los turnos, suspendía por
    #: «ofrecer una S que no está en la leyenda». Un gate que suspende por lo
    #: que no mide deja de leerse — el mismo razonamiento que ya excluyó el
    #: `"✓,✗,—"` del 04.
    columnas = set()
    for e in DV_CODIGOS.get(fname, []):
        c0 = column_index_from_string(e[1])
        c1 = column_index_from_string(e[2]) if e[2] else 16384
        columnas |= set((e[0], c) for c in range(c0, min(c1, 60) + 1))
    for ws in wb.worksheets:
        if ws.title not in hojas:
            continue
        for dv in ws.data_validations.dataValidation:
            f1 = getattr(dv, 'formula1', None)
            if not (isinstance(f1, str) and f1.startswith('"')):
                continue
            if columnas and not any(
                    (ws.title, c) in columnas
                    for c in _columnas_de_sqref(dv.sqref)):
                continue          # la DV no está en la rejilla de códigos
            valores = [v for v in f1.strip('"').split(',') if v]
            if not valores or max(len(v) for v in valores) > 2:
                continue          # no es una lista de códigos
            for v in valores:
                if v not in validas:
                    fallos.append('{}!{}: la DV ofrece «{}», que no está en la '
                                  'leyenda única'
                                  .format(ws.title, dv.sqref, v))
    return fallos


#: Los 8 valores de error de Excel, en las dos grafías (la inglesa que graba el
#: XML y la española que enseña la interfaz). Se comparan POR LISTA y no por
#: «empieza por #»: la cabecera de la columna de numeración del 04 y del 06 es
#: literalmente «#» (`04-onboarding-nuevo-empleado.xlsx:Checklist
#: Onboarding:A6`, `06-evaluacion-desempeno.xlsx:Ficha Evaluación:A11`), así
#: que el gate ingenuo suspendía por 5 cabeceras perfectamente sanas y escondía
#: los 3 errores reales entre ellas.
ERRORES_EXCEL = frozenset({
    '#DIV/0!', '#¡DIV/0!', '#N/A', '#N/D', '#NAME?', '#¿NOMBRE?',
    '#NULL!', '#¡NULO!', '#NUM!', '#¡NUM!', '#REF!', '#¡REF!',
    '#VALUE!', '#¡VALOR!', '#GETTING_DATA', '#SPILL!', '#¡DESBORDAMIENTO!',
    '#CALC!', '#¡CALC!',
})


def cacheados_con_error(path):
    """§1.5 — gate DURO: tras `inject_cache`, ninguna celda puede tener CACHEADO
    un valor de error de Excel.

    Es el gate que caza el `#DIV/0!` de `06-evaluacion-desempeno.xlsx:Ficha
    Evaluación:C22` —y el que arrastra a `C23`, el veredicto— en un documento
    que se firma delante del empleado (DOM-02/COM-04). Se le pasa la RUTA porque
    hay que abrir el libro con `data_only=True`.
    """
    import openpyxl
    fuera = []
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip().upper() in ERRORES_EXCEL:
                    fuera.append('{}!{}: valor cacheado «{}»'
                                 .format(ws.title, c.coordinate, v))
    return fuera


def contar(wb):
    """Censo interno de un libro: lo que main.py vuelca al informe."""
    r = dict(hojas=len(wb.sheetnames), formulas=0, verdes=0,
             verdes_con_dato=0, dv=0, cf=0, protegidas=0, celdas=0)
    for ws in wb.worksheets:
        r['dv'] += len(ws.data_validations.dataValidation)
        r['cf'] += sum(len(cf.rules) for cf in ws.conditional_formatting)
        if ws.protection.sheet:
            r['protegidas'] += 1
        for row in ws.iter_rows():
            for c in row:
                # el verde se cuenta AUNQUE la celda esté vacía: una casilla
                # editable en blanco es precisamente lo que se le ofrece al
                # cliente.
                if es_verde(c):
                    r['verdes'] += 1
                    if c.value is not None:
                        r['verdes_con_dato'] += 1
                if c.value is None:
                    continue
                r['celdas'] += 1
                if isinstance(c.value, str) and c.value.startswith('='):
                    r['formulas'] += 1
    return r


# ==========================================================================
# Entrada del pipeline
# ==========================================================================
def aplicar(wb, fname, informe):
    """PRE-pase del motor (§1), antes de que el grupo escriba su contenido.

    Aquí sólo va lo que NO depende de que los grupos hayan movido columnas: los
    colores de sección del 04 (que además desactivan la colisión con el verde de
    edición, y tienen que correr ANTES de `aplicar_verde`). Todo lo demás vive
    en `cerrar()`.
    """
    colores_seccion_04(wb, fname, informe)
    return informe


def cerrar(wb, fname, informe, proteger_hojas=True):
    """POST-pase del motor: lo que depende del layout FINAL.

    Devuelve el dict de gates del fichero (pendientes de grupo, pestañas citadas
    inexistentes, citas legales obsoletas, contadores).
    """
    pendientes = []

    # limpieza previa: la cadena vacía es un defecto del censo (`empty_str`)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    c.value = None

    escribir_leyenda(wb, fname, informe)
    aplicar_formatos(wb, fname, informe)
    aplicar_verde(wb, fname, informe)
    n_dv = aplicar_dv(wb, fname, informe, pendientes)
    n_cf = aplicar_cf(wb, fname, informe, pendientes)

    # ⚠ ORDEN: la bio y la versión AÑADEN dos filas a `Instrucciones`, así que
    # tienen que escribirse ANTES de fijar el `print_area`. Al revés, la 1.ª
    # pasada grababa `A1:B22` y la 2.ª —con las filas ya puestas— `A1:B24`: 9
    # diferencias de idempotencia, una por fichero, y ninguna visible en el
    # contenido. Medido el 2026-08-24.
    bio_y_version(wb, informe)
    set_metadata(wb, fname, informe)

    presentacion = dict((e[0], e) for e in PRESENTACION.get(fname, []))
    for ws in wb.worksheets:
        if ws.title == 'Instrucciones':
            print_setup(ws, None, landscape=False)
        else:
            p = presentacion.get(ws.title)
            if p:
                _hoja, freeze, titles, tcols, land, fitw = p
                hdr = int(titles.split(':')[0].strip('$')) if titles else None
                print_setup(ws, hdr, landscape=land, fit_to_width=fitw,
                            title_cols=tcols)
                if titles:
                    ws.print_title_rows = titles
                if freeze:
                    ws.freeze_panes = freeze
            else:
                bl = _bloques(fname, ws.title)
                print_setup(ws, bl[0][1] if bl else None,
                            landscape=ws.max_column >= 6)
        ultima = get_column_letter(max(1, ws.max_column))
        ws.print_area = 'A1:{}{}'.format(ultima, max(1, ws.max_row))
        if proteger_hojas:
            proteger(ws, informe)

    citadas = pestanas_citadas(wb)
    return {
        'fichero': fname,
        'dv_aplicadas': n_dv,
        'cf_aplicadas': n_cf,
        'pendientes_de_grupo': pendientes,
        'pestanas_citadas_inexistentes': ['Instrucciones!{}: «{}»'
                                          .format(f, n) for f, n in citadas],
        'citas_legales_obsoletas': citas_obsoletas(wb),
        'leyenda_incoherente': leyenda_coherente(wb, fname),
        'hojas_esqueleto': hojas_esqueleto(wb, fname),
        'contadores': contar(wb),
    }
