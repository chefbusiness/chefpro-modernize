#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
grupo_c.py — Kit de Gestión de Personal y Turnos v2.0

Reparto del orquestador (2026-08-24): este grupo construye TRES ficheros —

  · `06-evaluacion-desempeno.xlsx`               (SPEC §4, bloque «06»)
  · `07-directorio-plantilla.xlsx`               (SPEC §4, bloque «07»)
  · `BONUS-02-calculadora-plantilla-optima.xlsx` (SPEC §3, bloque «BONUS-02»)

El 04 y el 05 —que la SPEC mete en el mismo §4— los construye `grupo_b`; el
BONUS-02, que la SPEC empareja con el 03 dentro del §3, cae aquí porque es la
otra mitad del MISMO modelo de dimensionamiento: `main.demo_fte` exige que
`03!Previsión por Servicio` y `BONUS-02!Calculadora` devuelvan los mismos
**7 FTE** con los datos por defecto, así que la cadena de esta hoja se escribe
celda a celda igual que la del 03 (`grupo_b._n03_prevision`), con los mismos
rótulos y los mismos parámetros en verde.

Contrato con `main.py`: se exponen `FICHEROS`, `pre(wb, fname, cambios)` y
`demos(carpeta, origen)`. Todo el trabajo va en `pre()`, es decir ANTES de
`motor.aplicar` y mucho antes de `motor.cerrar`, porque el motor fija formatos,
verde, DV, formato condicional, área de impresión y protección sobre el layout
FINAL: cuanto antes esté ese layout, menos casos especiales.

Idempotencia (2.ª pasada = 0 diferencias): las hojas que cambian de GEOMETRÍA
—`07!Plantilla` pasa de 15 columnas a 22 y `07!Vencimientos` de 5 a 9— se
**reconstruyen enteras** en vez de parchearse; las DV propias se borran antes de
reescribirse (una `DataValidation` añadida dos veces son dos entradas distintas
en el `digest` de `main.py`). La zona que limpia `_instrucciones` se calcula a
partir del NÚMERO DE LÍNEAS y nunca alcanza al bloque de bio + versión que
`motor.bio_y_version` ancla debajo: si lo alcanzara, la 2.ª pasada lo
reescribiría dos filas más abajo y la idempotencia se rompería sin que nada
cambiara a la vista.
"""

import contextlib
import copy
import datetime
import os

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import motor

FICHEROS = [
    '06-evaluacion-desempeno.xlsx',
    '07-directorio-plantilla.xlsx',
    'BONUS-02-calculadora-plantilla-optima.xlsx',
]

F06, F07, FB2 = FICHEROS

#: Marca de las DV que escribe ESTE grupo. Empieza por `motor.MARCA_DV` a
#: propósito: así `motor._limpiar_dv(ws)` —que reconoce las suyas por ese
#: prefijo— también se lleva las mías cuando limpia una hoja, y no quedan dos
#: validaciones sobre la misma celda (Excel aplica la primera que encuentra).
MARCA_C = motor.MARCA_DV + 'c'

CAP = motor.CAPACIDAD                    # 30 empleados en toda hoja indexada

NEGRITA = Font(bold=True)
CENTRO = Alignment(horizontal='center', vertical='center')


# ==========================================================================
# Utilidades locales
# ==========================================================================
def _merges(ws, deseadas):
    """Fija EXACTAMENTE las combinaciones de la hoja. Idempotente por
    construcción: deshace todas y rehace la lista pedida."""
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for m in deseadas:
        ws.merge_cells(m)


def _anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


def _vaciar(ws):
    """Vacía valor, relleno y formato de las celdas que YA EXISTEN.

    Se recorre `iter_rows()` y no `ws.cell(r, c)`: pedir una celda que no existe
    la CREA y con ella agranda `ws.max_row`/`max_column`, y el `print_area` que
    fija `motor.cerrar` saldría con filas y columnas en blanco dentro. Y se
    limpia el RELLENO, no sólo el valor: la v1.1 pinta de verde columnas que
    aquí cambian de sitio o pasan a calculadas, y ese verde huérfano le diría al
    cliente que puede sobrescribir una fórmula.
    """
    for fila in ws.iter_rows():
        for cel in fila:
            if cel.value is not None:
                cel.value = None
            cel.fill = PatternFill()
            cel.number_format = 'General'
            cel.font = Font()
            cel.alignment = Alignment()
    for r in list(ws.row_dimensions):
        ws.row_dimensions[r].height = None


def _rehacer(wb, titulo, indice):
    """Reconstruye una hoja entera en su posición. Es la forma más barata de ser
    idempotente cuando cambia la GEOMETRÍA."""
    if titulo in wb.sheetnames:
        wb.remove(wb[titulo])
    return wb.create_sheet(titulo, indice)


def _rotulo(ws, fila, texto, col=1, negrita=False, italica=False, tam=None):
    cel = ws.cell(row=fila, column=col, value=texto)
    if negrita or italica or tam:
        cel.font = Font(bold=bool(negrita), italic=bool(italica),
                        size=tam or 11)
    return cel


def _f(ws, coord, formula, fmt=None):
    """Escribe una fórmula y la REGISTRA para que `main.py` compruebe con
    `data_only` que quedó con valor cacheado (o que pycel confirma que vale la
    cadena vacía). Sin el registro, una fórmula del grupo se queda fuera del
    gate `data_only_formulas_nuevas` y nadie se entera de que el cliente la ve
    en blanco en el visor del móvil."""
    cel = ws[coord]
    cel.value = formula
    if fmt:
        cel.number_format = fmt
    motor._reg(ws, coord, formula)
    return cel


def _verde(ws, coord, valor=None, fmt=None):
    cel = ws[coord]
    if valor is not None:
        cel.value = valor
    if fmt:
        cel.number_format = fmt
    motor.marcar_verde(ws, coord)
    return cel


def _dv(ws, rango, valores, titulo, prompt, error):
    """DV de lista marcada para poder retirarla en la 2.ª pasada."""
    formula = '"{}"'.format(','.join(valores))
    if len(formula) > 255:
        raise ValueError('DV inline de {} caracteres (>255): {}'
                         .format(len(formula), titulo))
    dv = DataValidation(
        type='list', formula1=formula, allow_blank=True,
        showErrorMessage=True, errorTitle=titulo, error=error,
        errorStyle='stop', showInputMessage=True,
        promptTitle='{} · {}'.format(MARCA_C, titulo), prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(rango)
    return dv


def _nota_celda(ws, coord, titulo, prompt):
    """Mensaje de entrada (sin validación) sobre una celda de parámetro: es
    donde vive el «esto lo fija tu convenio, edítalo» del §1.4."""
    dv = DataValidation(type=None, showInputMessage=True,
                        promptTitle='{} · {}'.format(MARCA_C, titulo),
                        prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(coord)
    return dv


def _cabecera(ws, fila, textos, col0=1):
    """Cabecera oscura, del mismo estilo que ya usa la v1.1."""
    for i, txt in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i, value=txt)
        cel.fill = PatternFill('solid', fgColor=motor.CAB)
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
    ws.row_dimensions[fila].height = 30


def _seccion(ws, fila, texto, ancho, color='ECEFF1'):
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font = Font(bold=True, size=11)
    cel.fill = PatternFill('solid', fgColor=color)
    for c in range(2, ancho + 1):
        ws.cell(row=fila, column=c).fill = PatternFill('solid', fgColor=color)


def _titulo(ws, ancho, titulo):
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'AI Chef Pro · aichef.pro — Kit Gestión de Personal y Turnos'
    ws['A2'].font = Font(size=9, color='666666')
    return ['A1:{}1'.format(get_column_letter(ancho)),
            'A2:{}2'.format(get_column_letter(ancho))]


def _pie(ws, fila, ancho):
    cel = ws.cell(row=fila, column=1, value=motor.PIE)
    cel.font = Font(size=8, color='888888')
    cel.alignment = Alignment(horizontal='center')
    return 'A{f}:{c}{f}'.format(f=fila, c=get_column_letter(ancho))


def _nota(ws, fila, texto, ancho=None, alto=None):
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font = Font(size=9, italic=True)
    cel.alignment = Alignment(wrap_text=True, vertical='top')
    if alto:
        ws.row_dimensions[fila].height = alto
    if ancho:
        return 'A{f}:{c}{f}'.format(f=fila, c=get_column_letter(ancho))
    return None


#: Zona MÍNIMA que limpia `_instrucciones`. Las tres hojas de Instrucciones que
#: toca este grupo llegan hoy a la fila 21 como mucho (`06!B21` es la línea de
#: versión 1.1), así que cualquier zona por debajo de 22 dejaría texto de la
#: v1.1 vivo por debajo del texto nuevo.
ZONA_MIN = 22


def _instrucciones(ws, lineas):
    """Reescribe el cuerpo de `Instrucciones` (columna B en este kit).

    ⚠ La zona que se limpia es EXACTAMENTE `len(lineas) + 1` filas y nunca
    `ws.max_row`: `motor.bio_y_version` ancla su bloque de tres líneas debajo y
    lo localiza por la nota de desprotección. Si la limpieza lo borrara, la 2.ª
    pasada volvería a colocarlo en `ws.max_row + 2` —dos filas MÁS ABAJO, porque
    las celdas de la 1.ª pasada siguen existiendo aunque valgan `None`— y la
    idempotencia se rompería con tres diferencias por fichero que no se ven en
    el contenido.
    """
    if len(lineas) + 1 < ZONA_MIN:
        lineas = list(lineas) + [None] * (ZONA_MIN - 1 - len(lineas))
    col = motor.col_instrucciones(ws)
    zona = len(lineas) + 1
    for r in range(2, zona + 1):
        ws.cell(row=r, column=col).value = None
    fila = 2
    for texto in lineas:
        if texto is None:
            fila += 1
            continue
        cel = ws.cell(row=fila, column=col, value=texto)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
        cel.font = Font(bold=True, size=13) if fila == 2 else Font()
        fila += 1
    return zona


# ==========================================================================
# 06 · EVALUACIÓN DE DESEMPEÑO  (SPEC §4)
# ==========================================================================
#: Las 10 competencias de la v1.1, intactas: son las que el cliente ya conoce y
#: las que la landing enumera («10 competencias clave para hostelería»). Lo que
#: cambia es la ESCALA (entra `N/A`) y la media, que ahora sale sólo de lo
#: valorado — a un cocinero de partida no se le puntúa «Atención al cliente»
#: (DOM-30).
COMPETENCIAS = [
    'Puntualidad y asistencia',
    'Presentación personal e higiene',
    'Trabajo en equipo y colaboración',
    'Iniciativa y proactividad',
    'Limpieza y orden del puesto',
    'Rapidez y eficiencia',
    'Atención al cliente',
    'Conocimiento de la carta / producto',
    'Gestión del estrés en servicio',
    'Comunicación con compañeros y superiores',
]

#: Escala de la DV. Deja de ser `decimal 1..5` (que no admite texto) para que
#: «no aplica» sea una opción explícita y no un hueco indistinguible de un
#: olvido: `COUNT` ignora el texto, así que `N/A` sale de la media sola.
ESCALA = ['1', '2', '3', '4', '5', 'N/A']

FICHA_ANCHO = 4
#: Geometría de la ficha. `motor.BLOQUES` declara el bloque de competencias en
#: `(hoja, 11, 12, 21, 21)`, así que la cabecera va en la 11 y las diez
#: competencias en 12..21 — moverlas dejaría al motor pintando otra cosa.
FICHA_HDR = 11
FICHA_R0 = 12
FICHA_R1 = 21
#: RD-24 · mínimo de competencias PUNTUADAS (número, las N/A no cuentan)
#: para que la ficha emita un nivel. Con menos, C23 pide que se puntúe.
MIN_COMPETENCIAS = 5

#: Bloques de texto libre: `(fila_del_titulo, título, primera_fila, filas)`.
FICHA_TEXTOS = [
    (26, 'FORTALEZAS DESTACADAS', 27, 3),
    (31, 'ÁREAS DE MEJORA', 32, 3),
    (36, 'OBJETIVOS PARA EL PRÓXIMO PERIODO', 37, 3),
]
FICHA_PLAN = 41          # título del PLAN DE DESARROLLO INDIVIDUAL
FICHA_FIRMAS = 47
FICHA_PIE = 50

#: Ejemplo REAL de oficio para la hoja «Ficha (ejemplo relleno)» (§7-bis.5).
#: Una jefa de partida de cuarto frío a la que NO se le puntúa «Atención al
#: cliente»: es el caso que justifica el `N/A` y el que hace visible que la
#: media sale de nueve competencias, no de diez.
EJEMPLO_FICHA = {
    'cabecera': [
        'Marta Ruiz Beltrán',
        'Jefa de partida — cuarto frío',
        'Q3 2026 (julio-septiembre)',
        'John Guerrero — jefe de cocina',
        datetime.datetime(2026, 9, 30),
    ],
    'notas': [
        (5, 'Ni un retraso en el trimestre; avisó con 48 h del único cambio '
            'de turno que pidió.'),
        (5, 'Uniforme y aseo impecables; corrige a los ayudantes cuando se '
            'saltan el gorro.'),
        (4, 'Se lleva bien con pase y con sala; en el cambio de carta tiró '
            'del equipo.'),
        (4, 'Propuso el escandallo del steak tartar y bajó la merma del '
            'salmón un 6 %.'),
        (5, 'Cuarto frío cierra siempre en orden; los registros APPCC al día.'),
        (4, 'Va sobrada en producción; se atasca cuando entran dos comandas '
            'de tartar seguidas.'),
        ('N/A', 'No aplica: no atiende sala en este puesto.'),
        (4, 'Domina su partida; le falta soltura con la bodega y los postres.'),
        (3, 'En los servicios de más de 90 cubiertos pierde el orden de los '
            'pases. Es el punto a trabajar.'),
        (4, 'Pregunta lo que no sabe y da parte de las incidencias por '
            'escrito.'),
    ],
    'textos': [
        ['Organización y limpieza del cuarto frío: la partida mejor llevada '
         'del equipo.',
         'Muy fiable con el producto: cero mermas por mala conservación en '
         'todo el trimestre.',
         'Formadora natural — los dos ayudantes nuevos aprendieron con ella.'],
        ['Gestión del estrés cuando se juntan dos comandas de tartar en el '
         'mismo pase.',
         'Conocimiento de bodega y de la partida de postres, que hoy no '
         'cubre.',
         ''],
        ['Cubrir la partida de postres dos servicios por semana a partir de '
         'noviembre.',
         'Sacar adelante 4 servicios de más de 90 cubiertos sin retrasos en '
         'el pase.',
         ''],
    ],
    'plan': [
        ('Curso interno de organización de partida en servicio alto (8 h, '
         'dos sesiones)', datetime.datetime(2026, 11, 15),
         'Jefe de cocina · 4 servicios de +90 cubiertos sin retraso en el '
         'pase'),
        ('Rotación de dos servicios semanales en la partida de postres',
         datetime.datetime(2026, 12, 20),
         'Jefa de pastelería · saca la carta de postres sin apoyo'),
        ('', None, ''),
    ],
}


def _ficha(ws, ejemplo=None):
    """Construye la ficha de evaluación (en blanco o rellena con el ejemplo).

    Lo que había: `C22='=AVERAGE(C12:C21)'` sobre diez celdas vacías, con el
    valor **cacheado** `#DIV/0!` en el fichero LIVE, y `C23` encadenando el
    error. Es el documento que se imprime y se FIRMA delante del empleado
    (DOM-02/TEC-03/COM-04), y era el único `#ERROR` cacheado de los 9 ficheros.
    """
    _merges(ws, [])
    _vaciar(ws)
    _anchos(ws, {'A': 5, 'B': 35, 'C': 16, 'D': 40})
    merges = _titulo(ws, FICHA_ANCHO, 'Ficha de Evaluación de Desempeño'
                     + (' — EJEMPLO RELLENO' if ejemplo else ''))

    campos = ['Empleado:', 'Puesto:', 'Periodo evaluado:', 'Evaluador:',
              'Fecha evaluación:']
    for i, txt in enumerate(campos):
        fila = 4 + i
        _rotulo(ws, fila, txt, col=2, negrita=True)
        ref = 'C{}:D{}'.format(fila, fila)
        merges.append(ref)
        cel = ws['C{}'.format(fila)]
        if ejemplo:
            cel.value = ejemplo['cabecera'][i]
        if txt.startswith('Fecha'):
            cel.number_format = motor.FMT_FECHA
        motor.marcar_verde(ws, ref)

    _rotulo(ws, 10, 'COMPETENCIAS — puntúa de 1 a 5, o N/A si la competencia '
                    'no aplica al puesto', negrita=True)
    ws['A10'].fill = PatternFill('solid', fgColor='ECEFF1')
    for c in range(2, FICHA_ANCHO + 1):
        ws.cell(row=10, column=c).fill = PatternFill('solid', fgColor='ECEFF1')
    merges.append('A10:D10')

    _cabecera(ws, FICHA_HDR, ['#', 'Competencia', 'Puntuación (1-5)',
                              'Observaciones'])
    for i, nombre in enumerate(COMPETENCIAS):
        fila = FICHA_R0 + i
        ws.cell(row=fila, column=1, value=i + 1).alignment = CENTRO
        cel = ws.cell(row=fila, column=2, value=nombre)
        cel.alignment = Alignment(wrap_text=True, vertical='center')
        if ejemplo:
            nota, texto = ejemplo['notas'][i]
            ws.cell(row=fila, column=3, value=nota).alignment = CENTRO
            obs = ws.cell(row=fila, column=4, value=texto)
            obs.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[fila].height = 26

    # ⚠ La DV heredada es `whole` 1..5 sobre C12:C21 y NO admite texto: con ella
    # puesta, «N/A» se rechazaría. Se retira por SOLAPE de rango, que es lo
    # único que la reconoce (no tiene `promptTitle`).
    motor._limpiar_dv(ws, motor._celdas_sqref('C12:C21'))
    _dv(ws, 'C{}:C{}'.format(FICHA_R0, FICHA_R1), ESCALA,
        'Puntuación no válida',
        '1 = Deficiente · 2 = Mejorable · 3 = Adecuado · 4 = Bueno · '
        '5 = Excelente. Elige N/A cuando la competencia no aplique al puesto '
        '(un cocinero de partida no atiende sala): la media se calcula SÓLO '
        'con lo valorado, así que N/A no penaliza.',
        'Escribe 1, 2, 3, 4, 5 o N/A.')

    _rotulo(ws, 22, 'PUNTUACIÓN MEDIA', col=2, negrita=True)
    _f(ws, 'C22', motor.guarda_media('$C${}:$C${}'.format(FICHA_R0, FICHA_R1)),
       motor.FMT_DEC2)
    _rotulo(ws, 23, 'NIVEL', col=2, negrita=True)
    # RD-24 · el nivel se emitía con UNA sola competencia puntuada: con `C12=5`
    # y las otras nueve en blanco, `C22=5,00` y `C23='⭐ EXCELENTE'` en el
    # documento que se firma y que puede acabar en un expediente. Es el mismo
    # «veredicto sin datos suficientes» que se corrigió en el semáforo del 03.
    # El contador de C24 ya decía «1 de 10», pero es una celda que nadie mira
    # antes de firmar: el aviso tiene que estar EN la casilla del nivel, y con
    # el mismo ámbar del semáforo (motor.VOC_NIVEL).
    _f(ws, 'C23',
       '=IF(COUNT($C${r0}:$C${r1})<{m},'
       '"⚠ puntúa al menos {m} competencias",'
       'IF($C$22="","",'
       'IF($C$22>=4.5,"⭐ EXCELENTE",'
       'IF($C$22>=3.5,"✓ BUENO",'
       'IF($C$22>=2.5,"→ ADECUADO",'
       'IF($C$22>=1.5,"⚠ MEJORABLE","✗ DEFICIENTE"))))))'
       .format(r0=FICHA_R0, r1=FICHA_R1, m=MIN_COMPETENCIAS))
    ws['C23'].alignment = Alignment(vertical='center')
    if ejemplo:
        # `motor.CF_CELDA` sólo cubre 'Ficha Evaluación'; la hoja de ejemplo es
        # nueva y nadie la colorearía.
        _cf_propio(ws, 'C23:C23', motor.VOC_NIVEL)
    _rotulo(ws, 24, 'Competencias valoradas', col=2)
    _f(ws, 'C24', '=COUNT($C${}:$C${})&" de {}"'
       .format(FICHA_R0, FICHA_R1, len(COMPETENCIAS)))
    ws['C24'].alignment = CENTRO
    ws['C24'].font = Font(size=9, color='666666')
    ws['D24'] = ('← las N/A no cuentan: la media sale sólo de las '
                 'competencias puntuadas')
    ws['D24'].font = Font(size=9, italic=True, color='666666')

    for i, bloque in enumerate(FICHA_TEXTOS):
        r_tit, titulo, r0, n = bloque
        _seccion(ws, r_tit, titulo, FICHA_ANCHO)
        merges.append('A{f}:D{f}'.format(f=r_tit))
        for j in range(n):
            fila = r0 + j
            ws.cell(row=fila, column=1, value='▸').alignment = CENTRO
            ref = 'B{f}:D{f}'.format(f=fila)
            merges.append(ref)
            if ejemplo:
                ws.cell(row=fila, column=2,
                        value=ejemplo['textos'][i][j] or None)
            motor.marcar_verde(ws, ref)
            for c in range(2, FICHA_ANCHO + 1):
                ws.cell(row=fila, column=c).alignment = Alignment(
                    wrap_text=True, vertical='top')
            ws.row_dimensions[fila].height = 30

    _seccion(ws, FICHA_PLAN, 'PLAN DE DESARROLLO INDIVIDUAL', FICHA_ANCHO)
    merges.append('A{f}:D{f}'.format(f=FICHA_PLAN))
    _cabecera(ws, FICHA_PLAN + 1,
              ['#', 'Acción formativa', 'Fecha objetivo',
               'Responsable e indicador de logro'])
    for j in range(3):
        fila = FICHA_PLAN + 2 + j
        ws.cell(row=fila, column=1, value=j + 1).alignment = CENTRO
        if ejemplo:
            accion, fecha, indicador = ejemplo['plan'][j]
            ws.cell(row=fila, column=2, value=accion or None)
            ws.cell(row=fila, column=3, value=fecha)
            ws.cell(row=fila, column=4, value=indicador or None)
        ws.cell(row=fila, column=3).number_format = motor.FMT_FECHA
        motor.marcar_verde(ws, 'B{f}:D{f}'.format(f=fila))
        for c in range(2, FICHA_ANCHO + 1):
            ws.cell(row=fila, column=c).alignment = Alignment(
                wrap_text=True, vertical='top')
        ws.row_dimensions[fila].height = 30

    _rotulo(ws, FICHA_FIRMAS,
            'Firma evaluador: _________________          '
            'Firma empleado: _________________')
    merges.append('A{f}:D{f}'.format(f=FICHA_FIRMAS))
    _rotulo(ws, FICHA_FIRMAS + 1, 'Fecha: ___/___/______')
    merges.append('A{f}:D{f}'.format(f=FICHA_FIRMAS + 1))
    merges.append(_pie(ws, FICHA_PIE, FICHA_ANCHO))
    _merges(ws, merges)


def _tendencia(fila):
    """RD-05/RT-10 — los dos últimos trimestres INFORMADOS, con huecos.

    La versión anterior usaba `INDEX($B5:$E5, COUNT($B5:$E5))`, y ahí está el
    error: `COUNT` cuenta CUÁNTAS celdas tienen número e `INDEX` toma una
    POSICIÓN. Sólo coinciden si los trimestres se rellenan de izquierda a
    derecha sin saltarse ninguno — que es justo lo contrario del caso que el
    hallazgo original quería cubrir: quien evalúa dos veces al año deja huecos
    en medio.

    Medido en la copia dry-run: con `Q1=4,0 · Q2 vacío · Q3=4,5`, `COUNT` da 2,
    `INDEX(…,2)` apunta a Q2 —vacío, que vale 0— y la hoja escribía «↓ Baja»
    sobre una mejora de medio punto. Y con `Q1=5 · Q2=1 · Q3 vacío · Q4=4`,
    «↓ Baja» sobre una subida de tres puntos.

    Se sustituye por dos cascadas de `IF` sobre E, D, C y B —el último
    informado y el anterior a ése—, que no dependen de la posición y usan sólo
    comparaciones, verificadas con pycel. Nada de `LOOKUP`, que la SPEC no
    tiene medido.
    """
    ult = ('IF($E{f}<>"",$E{f},IF($D{f}<>"",$D{f},'
           'IF($C{f}<>"",$C{f},$B{f})))'.format(f=fila))
    pen = ('IF($E{f}<>"",IF($D{f}<>"",$D{f},IF($C{f}<>"",$C{f},$B{f})),'
           'IF($D{f}<>"",IF($C{f}<>"",$C{f},$B{f}),'
           'IF($C{f}<>"",$B{f},"")))'.format(f=fila))
    return ('=IF(COUNT($B{f}:$E{f})<2,"",IF({u}>{p},"↑ Mejora",'
            'IF({u}<{p},"↓ Baja","→ Estable")))'
            .format(f=fila, u=ult, p=pen))


def _n06_historico(wb, cambios):
    """§4 · TEC-27/COM-21/DOM-30 — la tendencia compara los DOS ÚLTIMOS
    trimestres informados.

    Lo que había: `G5='=IF(AND(E5<>"",D5<>""),IF(E5>D5,…))'`, que sólo mira Q4
    contra Q3 y por tanto está en blanco NUEVE MESES al año en la columna que
    justifica la hoja. Y sólo 15 filas para los 30 empleados del kit (DOM-32).
    """
    ws = wb['Histórico']
    motor._limpiar_dv(ws)
    _merges(ws, [])
    _vaciar(ws)
    _anchos(ws, {'A': 26, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 14,
                 'G': 20})
    merges = _titulo(ws, 7, 'Histórico de Evaluaciones')
    _cabecera(ws, 4, ['Empleado', 'Q1', 'Q2', 'Q3', 'Q4', 'Media Anual',
                      'Tendencia'])
    r0, r1 = 5, 4 + CAP
    for fila in range(r0, r1 + 1):
        _f(ws, 'F{}'.format(fila),
           motor.guarda_media('$B{f}:$E{f}'.format(f=fila)), motor.FMT_DEC2)
        _f(ws, 'G{}'.format(fila), _tendencia(fila))
        for c in range(2, 6):
            ws.cell(row=fila, column=c).number_format = motor.FMT_DEC1
            ws.cell(row=fila, column=c).alignment = CENTRO
        ws.cell(row=fila, column=7).alignment = CENTRO

    nota0 = r1 + 2
    _nota(ws, nota0,
          '▸ Vuelca aquí la PUNTUACIÓN MEDIA (celda C22) de cada ficha: una '
          'columna por trimestre y una fila por empleado. La Tendencia '
          'compara los DOS ÚLTIMOS trimestres informados aunque haya huecos '
          'en medio: si evalúas dos veces al año, deja los trimestres que no '
          'toques en blanco y la comparación sigue siendo la correcta.', 7, 30)
    _nota(ws, nota0 + 1,
          '▸ «Media Anual» se queda en blanco mientras no haya ni una nota: '
          'un empleado sin evaluar no es un empleado con un cero.', 7, 15)
    _nota(ws, nota0 + 2,
          '▸ La hoja admite 30 empleados, los mismos que el directorio (07) y '
          'que el cuadrante (01).', 7, 15)
    merges += ['A{f}:G{f}'.format(f=nota0 + i) for i in range(3)]
    merges.append(_pie(ws, nota0 + 4, 7))
    _merges(ws, merges)
    cambios.append('06:Histórico!G5:G34: la tendencia compara los DOS últimos '
                   'trimestres INFORMADOS aunque falte uno intermedio — el '
                   'INDEX/COUNT anterior confundía posición con recuento y '
                   'devolvía el veredicto contrario (Q1=4,0 · Q2 vacío · '
                   'Q3=4,5 daba «↓ Baja»); la hoja pasa de 15 a 30 empleados '
                   '(TEC-27/COM-21/DOM-32/RD-05/RT-10)')


def _n06_instrucciones(wb, cambios):
    ws = wb['Instrucciones']
    motor._limpiar_dv(ws)
    _instrucciones(ws, [
        '06 · Evaluación de Desempeño',
        None,
        'Cómo usar esta plantilla:',
        "▸ La hoja 'Ficha Evaluación' viene EN BLANCO: es el original. Para "
        'evaluar a alguien, duplícala (clic derecho sobre la pestaña → Mover o '
        'copiar → Crear una copia) y renombra la copia «Apellido-Q3», por '
        'ejemplo.',
        "▸ La hoja 'Ficha (ejemplo relleno)' es esa misma ficha con un caso "
        'real terminado. No la borres: es la referencia de cómo se rellenan '
        'las observaciones y el plan de desarrollo.',
        '▸ Puntúa cada competencia de 1 a 5, o elige N/A si no aplica al '
        'puesto. La media sale SÓLO de lo valorado y la casilla «Competencias '
        'valoradas» te dice sobre cuántas se ha calculado.',
        "▸ Cuando cierres una ficha, copia su PUNTUACIÓN MEDIA a la hoja "
        "'Histórico', en la columna del trimestre que toque.",
        None,
        'Escala de puntuación:',
        '▸ 1 = Deficiente — No cumple los requisitos mínimos del puesto',
        '▸ 2 = Mejorable — Cumple parcialmente, necesita supervisión constante',
        '▸ 3 = Adecuado — Cumple lo esperado para el puesto',
        '▸ 4 = Bueno — Supera expectativas en varias áreas',
        '▸ 5 = Excelente — Referencia para el equipo',
        '▸ N/A = No aplica al puesto. No es un cero: queda fuera de la media.',
        None,
        'Nota media y nivel:',
        '▸ Con la ficha recién abierta, la media y el nivel están EN BLANCO. '
        'Es lo correcto: un empleado sin puntuar no tiene nota, y la ficha se '
        'imprime y se firma tal cual.',
        '▸ El nivel se colorea solo: verde a partir de 3,5, ámbar entre 1,5 y '
        '3,5 y rojo por debajo.',
        None,
        'Histórico y tendencia:',
        '▸ La columna Tendencia compara los DOS ÚLTIMOS trimestres '
        'informados, no sólo Q4 contra Q3: con Q1 y Q2 rellenos ya dice algo.',
        '▸ Rellena los trimestres de izquierda a derecha (Q1, luego Q2…): la '
        'comparación cuenta cuántas notas hay, no en qué columna están.',
        None,
        'Plan de desarrollo individual:',
        '▸ El bloque del final de la ficha es lo que convierte una evaluación '
        'en una mejora: acción formativa, fecha objetivo y quién responde de '
        'que ocurra, con un indicador que se pueda comprobar.',
        '▸ Revísalo en la evaluación siguiente antes de puntuar nada.',
        None,
        'Frecuencia recomendada:',
        '▸ Evaluación formal cada 3 meses (trimestral).',
        '▸ Feedback informal continuo — al menos 1 conversación/semana.',
    ])
    cambios.append('06:Instrucciones: escala con N/A, cómo duplicar la ficha, '
                   'cómo volcar la media al histórico y qué es el plan de '
                   'desarrollo (COM-26/COM-21/DOM-30)')


def _n06(wb, cambios):
    _ficha(wb['Ficha Evaluación'])
    cambios.append('06:Ficha Evaluación!C22/C23/C24: media con guarda de rango '
                   'vacío, nivel que se calla sin media y contador de '
                   'competencias valoradas — la v1.1 entregaba «#DIV/0!» '
                   'CACHEADO en las dos celdas de un documento que se firma '
                   '(DOM-02/TEC-03/COM-04)')
    cambios.append('06:Ficha Evaluación!C12:C21: la escala pasa de decimal '
                   '1-5 a lista 1,2,3,4,5,N/A — a un cocinero de partida no se '
                   'le puntúa «Atención al cliente» y la media sale sólo de lo '
                   'valorado (DOM-30)')
    cambios.append('06:Ficha Evaluación!B27:D29/B32:D34/B37:D39: wrap_text y '
                   'alto de fila en las nueve celdas combinadas de texto '
                   'libre — sin ellos el comentario se corta al imprimir la '
                   'ficha que se firma (TEC-24)')
    cambios.append('06:Ficha Evaluación!A41:D45: bloque PLAN DE DESARROLLO '
                   'INDIVIDUAL (acción · fecha objetivo · responsable e '
                   'indicador), que la landing prometía y no existía (COM-21)')
    _ficha(_rehacer(wb, 'Ficha (ejemplo relleno)',
                    wb.sheetnames.index('Ficha Evaluación') + 1),
           ejemplo=EJEMPLO_FICHA)
    cambios.append('06:Ficha (ejemplo relleno): hoja nueva — la misma ficha '
                   'con un caso real terminado (9 competencias puntuadas y '
                   'una N/A), que es la respuesta a «¿una ficha o varias?» '
                   'de §7.5 (COM-26)')
    _n06_historico(wb, cambios)
    _n06_instrucciones(wb, cambios)


# ==========================================================================
# 07 · DIRECTORIO DE PLANTILLA  (SPEC §4)
# ==========================================================================
#: Las columnas del directorio: `(letra, cabecera, ancho, formato)`.
#:
#: ⚠ `D` sigue siendo «Tipo Contrato» y `G` sigue siendo «Jornada» porque
#: `motor.DV_LISTA['07-directorio-plantilla.xlsx']` las localiza POR LETRA con
#: un centinela de cabecera: moverlas dejaría las dos validaciones sin colocar
#: y el motor lo anotaría como «DV aplazada al grupo» — es decir, sin
#: desplegable de modalidad de contrato en el fichero que va a la gestoría.
#:
#: ⚠ «Alérgenos Propios» NO está: es un dato de salud, categoría especial del
#: art. 9 RGPD, y la v1.1 lo pedía al mismo nivel que la talla de camiseta. No
#: se sustituye por nada (DOM-13/COM-25).
#:
#: ⚠ La SPEC §4 dice «21 columnas A4:U4 … más una columna calculada Aviso».
#: Las dos cosas no caben: 15 columnas de la v1.1 − «Alérgenos Propios» + las 7
#: nuevas son ya 21, así que «Aviso» es la 22.ª y la hoja llega a `V`. Se
#: mantiene la SUSTANCIA (ninguna columna de la v1.1 se pierde salvo la de
#: salud, entran las siete que la landing enumeraba y el aviso de minoría de
#: edad es una columna calculada) y se anota la desviación de la letra.
COLS_07 = [
    ('A', 'Nombre Completo', 24, None),
    ('B', 'DNI / NIE', 14, '@'),
    ('C', 'Fecha de Nacimiento', 16, None),
    ('D', 'Tipo Contrato', 16, None),
    ('E', 'Puesto', 20, None),
    ('F', 'Grupo Profesional (convenio)', 20, None),
    # RT-20 · la DV de esta columna sólo admite «Completa» y «Parcial»,
    # y heredaba el '0.00' del bloque numérico vecino: un formato que
    # contradice lo que se puede escribir en la celda.
    ('G', 'Jornada', 12, 'General'),
    ('H', 'Horas Semanales', 12, None),
    ('I', 'Convenio Aplicable', 26, None),
    ('J', 'NAF / Nº Seguridad Social', 20, '@'),
    ('K', 'Fecha Alta', 13, None),
    ('L', 'Fecha Fin Contrato', 15, None),
    ('M', 'Fin Periodo de Prueba', 15, None),
    ('N', 'Caducidad Carnet Manipulador', 17, None),
    ('O', 'Caducidad PRL', 15, None),
    ('P', 'Salario Bruto Anual', 16, None),
    ('Q', 'Teléfono', 14, '@'),
    ('R', 'Email', 24, None),
    ('S', 'Contacto Emergencia', 24, None),
    ('T', 'Talla Uniforme', 12, None),
    ('U', 'Taquilla', 10, None),
    ('V', 'Aviso', 34, None),
]
ANCHO_07 = len(COLS_07)                       # 22 → A..V
COL_AVISO = 'V'
R0_07, R1_07 = 5, 4 + CAP                     # 5..34

#: Los cuatro vencimientos del §4: `(rótulo, columna de la Plantilla)`. El
#: orden manda: cada uno ocupa un par (fecha, alerta) en `Vencimientos`, y las
#: cuatro columnas de alerta caen en C, E, G e I, que es donde
#: `motor.CF_COLUMNA` espera encontrar la cabecera «Alerta».
VENCIMIENTOS = [
    ('Fin de contrato', 'L'),
    ('Fin periodo de prueba', 'M'),
    ('Caducidad carnet de manipulador', 'N'),
    ('Caducidad de la formación PRL', 'O'),
]
VENC_HDR = 6
VENC_R0, VENC_R1 = 7, 6 + CAP                 # 7..36
ANCHO_VENC = 1 + 2 * len(VENCIMIENTOS)        # 9 → A..I

#: RD-15/RT-07 · La mayoría de edad NO se cuenta en días. La versión anterior
#: usaba `DIAS_18 = 6570` (18 × 365) y el aviso se APAGABA entre cuatro y cinco
#: días ANTES del 18.º cumpleaños, porque entre 2008 y 2026 hay cinco bisiestos
#: (2008, 2012, 2016, 2020, 2024): 18 años reales son 6.574-6.575 días. Medido:
#: nacido el 28/08/2008 —aún menor el 24/08/2026— llevaba 6.570 días y el aviso
#: ya no salía. Justo en esos días sigue viva la prohibición de nocturnidad y
#: de horas extra del art. 6 ET.
#:
#: Se sustituye por la fecha exacta del cumpleaños con `DATE(YEAR+18, MONTH,
#: DAY)`, las tres verificadas con pycel en la cabecera de la SPEC.
FECHA_18 = 'DATE(YEAR($C{f})+18,MONTH($C{f}),DAY($C{f}))'


def _n07_plantilla(wb, cambios):
    """§4 · DOM-13/DOM-24/TEC-19/COM-16/COM-25 — el directorio que pide una
    gestoría, y sin datos de salud."""
    ws = wb['Plantilla']
    motor._limpiar_dv(ws)
    _merges(ws, [])
    _vaciar(ws)
    _anchos(ws, dict((c[0], c[2]) for c in COLS_07))
    merges = _titulo(ws, ANCHO_07, 'Directorio de Plantilla')
    _cabecera(ws, 4, [c[1] for c in COLS_07])
    for letra, _cab, _ancho, fmt in COLS_07:
        if not fmt:
            continue
        for fila in range(R0_07, R1_07 + 1):
            ws['{}{}'.format(letra, fila)].number_format = fmt
    for fila in range(R0_07, R1_07 + 1):
        # Aviso: lo ÚNICO calculado de la hoja. La fecha de nacimiento entró
        # por esto (DOM-24): sin ella, ni el cuadrante ni nadie puede saber que
        # a esa persona no se le puede poner un turno de noche.
        _f(ws, '{}{}'.format(COL_AVISO, fila),
           '=IF($C{f}="","",IF({d}>TODAY(),'
           '"⚠ MENOR DE EDAD: sin nocturnidad ni horas extra (art. 6 ET)",'
           '""))'.format(f=fila, d=FECHA_18.format(f=fila)))
        ws['{}{}'.format(COL_AVISO, fila)].alignment = Alignment(
            wrap_text=True, vertical='center')

    nota0 = R1_07 + 2
    _nota(ws, nota0,
          '▸ Esta hoja NO pide datos de salud. La versión 1.1 traía una '
          'columna con las alergias e intolerancias del trabajador y se ha '
          'retirado: son categoría especial del art. 9 RGPD y no se tratan al '
          'mismo nivel que la talla de camiseta. Lee la sección de protección '
          "de datos en 'Instrucciones'.",
          ANCHO_07, 30)
    _nota(ws, nota0 + 1,
          '▸ «Fin periodo de prueba», «Caducidad carnet manipulador» y '
          "«Caducidad PRL» alimentan solas la hoja 'Vencimientos': escríbelas "
          'aquí y no las teclees dos veces.', ANCHO_07, 30)
    _nota(ws, nota0 + 2,
          '▸ «Convenio aplicable» y «Grupo profesional» son los dos campos que '
          'pide primero una gestoría o una inspección; el NAF es el número de '
          'afiliación a la Seguridad Social del trabajador, no el de la '
          'empresa.', ANCHO_07, 30)
    _nota(ws, nota0 + 3,
          '▸ La columna «Aviso» avisa de la minoría de edad en cuanto pongas '
          'la fecha de nacimiento (art. 6 ET: sin trabajo nocturno ni horas '
          'extraordinarias).', ANCHO_07, 30)
    merges += ['A{f}:{c}{f}'.format(f=nota0 + i, c=get_column_letter(ANCHO_07))
               for i in range(4)]
    merges.append(_pie(ws, nota0 + 5, ANCHO_07))
    _merges(ws, merges)

    # RD-06 · el autofiltro venía de la v1.1 con `A4:O34` mientras la tabla
    # crecía a 22 columnas: ordenar desde el desplegable reordenaba A:O y
    # dejaba P:V clavadas — salario, teléfono, email, contacto de emergencia,
    # talla, taquilla y el aviso de menor de edad cruzados con OTRA persona, en
    # el fichero más sensible del kit. Y `Instrucciones!B6` invita
    # expresamente a usarlo. Se fija a la última columna real; el gate
    # `gate_autofiltro` de `main.py` impide que vuelva a quedarse corto.
    ws.auto_filter.ref = 'A4:{}{}'.format(get_column_letter(ANCHO_07), R1_07)
    cambios.append('07:Plantilla: auto_filter A4:O34 → A4:{}{} — abarca por '
                   'fin las 22 columnas; ordenar por el desplegable ya no '
                   'cruza salario, teléfono, contacto de emergencia y aviso '
                   'de menor de edad con la persona equivocada (RD-06)'
                   .format(get_column_letter(ANCHO_07), R1_07))
    cambios.append('07:Plantilla!A4:V4: 22 columnas — entran NAF, fecha de '
                   'nacimiento, convenio aplicable, grupo profesional, fin de '
                   'periodo de prueba, caducidad del carnet de manipulador y '
                   'caducidad de PRL, y SALE «Alérgenos Propios» (art. 9 '
                   'RGPD) (DOM-13/DOM-24/TEC-19/COM-16)')
    cambios.append('07:Plantilla!V5:V34: columna calculada «Aviso» con la '
                   'alerta de minoría de edad (art. 6 ET), que es para lo que '
                   'entra la fecha de nacimiento (DOM-24)')


def _n07_vencimientos(wb, cambios):
    """§4 · DOM-14/TEC-07 — los 30 empleados y los CUATRO vencimientos.

    Lo que había: 15 filas (`A7:E21` leyendo `Plantilla!A5:A19`) para una hoja
    que admite 30 personas, un solo vencimiento —el del contrato— y un bloque
    de carnets de entrada MANUAL que no leía nada del directorio. Del empleado
    16 en adelante nadie avisaba, y un temporal que vence sin preaviso se
    convierte en indefinido por ministerio de la ley.
    """
    ws = wb['Vencimientos']
    motor._limpiar_dv(ws)
    _merges(ws, [])
    _vaciar(ws)
    anchos = {'A': 26}
    for i in range(len(VENCIMIENTOS)):
        anchos[get_column_letter(2 + 2 * i)] = 17
        anchos[get_column_letter(3 + 2 * i)] = 20
    _anchos(ws, anchos)
    merges = _titulo(ws, ANCHO_VENC, 'Alertas de Vencimientos')

    _rotulo(ws, 3, 'Fecha de hoy:', negrita=True)
    _f(ws, 'B3', '=TODAY()', motor.FMT_FECHA)
    # RT-24 · es el ÚNICO punto del kit donde una función volátil se cruza con
    # el cache. En Excel no hay problema (los nueve libros llevan
    # `fullCalcOnLoad`), pero el cache existe justo para los visores que NO
    # recalculan —la vista previa de Drive, el Quick Look del móvil—, y ahí la
    # celda enseña para siempre la fecha en que se construyó el fichero y los
    # 120 semáforos que dependen de ella son los de ese día. No se puede dejar
    # sin cachear sin tocar `inject_cache.py`, que es compartido por todos los
    # kits; se avisa al lado de la celda, que es donde se lee.
    _nota_celda(ws, 'B3', 'Se actualiza al abrir en Excel',
                'Esta celda es =HOY(). Excel la recalcula al abrir el fichero. '
                'En la VISTA PREVIA del móvil o de Drive, que no recalcula, '
                'verás la fecha en que se descargó el kit y los semáforos de '
                'esa fecha: ábrelo en Excel o en LibreOffice antes de fiarte '
                'de los avisos.')
    _rotulo(ws, 4, 'Aviso ROJO — faltan menos de (días):')
    _verde(ws, 'B4', 30, motor.FMT_ENT)
    _nota_celda(ws, 'B4', 'Umbral del aviso rojo',
                'A partir de aquí el vencimiento es urgente. Treinta días es '
                'lo razonable para un contrato temporal (el preaviso del art. '
                '49.1.c ET son 15 días en los de más de un año) y de sobra '
                'para renovar un carnet de manipulador.')
    _rotulo(ws, 5, 'Aviso ÁMBAR — faltan menos de (días):')
    _verde(ws, 'B5', 60, motor.FMT_ENT)
    _nota_celda(ws, 'B5', 'Umbral del aviso ámbar',
                'La antelación con la que quieres enterarte. Súbelo si tu '
                'formación de PRL la imparte un tercero y tarda en darte '
                'fecha.')

    rango_alertas = []
    for i in range(len(VENCIMIENTOS)):
        rango_alertas.append('$' + get_column_letter(3 + 2 * i))
    _rotulo(ws, 3, 'Vencimientos ya VENCIDOS:', col=4, negrita=True)
    _f(ws, 'E3', '=' + '+'.join(
        'COUNTIF({c}${r0}:{c}${r1},"❌*")'
        .format(c=c, r0=VENC_R0, r1=VENC_R1) for c in rango_alertas),
       motor.FMT_ENT)
    _rotulo(ws, 4, 'En ROJO (vencen pronto):', col=4)
    _f(ws, 'E4', '=' + '+'.join(
        'COUNTIF({c}${r0}:{c}${r1},"🔴*")'
        .format(c=c, r0=VENC_R0, r1=VENC_R1) for c in rango_alertas),
       motor.FMT_ENT)
    _rotulo(ws, 5, 'En ÁMBAR (conviene mirarlo):', col=4)
    _f(ws, 'E5', '=' + '+'.join(
        'COUNTIF({c}${r0}:{c}${r1},"🟡*")'
        .format(c=c, r0=VENC_R0, r1=VENC_R1) for c in rango_alertas),
       motor.FMT_ENT)
    for r in (3, 4, 5):
        ws.cell(row=r, column=5).alignment = CENTRO

    cabeceras = ['Empleado']
    for rotulo, _col in VENCIMIENTOS:
        cabeceras += [rotulo, 'Alerta']
    _cabecera(ws, VENC_HDR, cabeceras)

    for j in range(CAP):
        fila = VENC_R0 + j
        origen = R0_07 + j
        _f(ws, 'A{}'.format(fila),
           '=IF(Plantilla!$A{o}="","",Plantilla!$A{o})'.format(o=origen))
        for i, par in enumerate(VENCIMIENTOS):
            col_fecha = get_column_letter(2 + 2 * i)
            col_alerta = get_column_letter(3 + 2 * i)
            _f(ws, '{c}{f}'.format(c=col_fecha, f=fila),
               '=IF(Plantilla!${p}{o}="","",Plantilla!${p}{o})'
               .format(p=par[1], o=origen), motor.FMT_FECHA)
            # Los dos umbrales viven en `B4` y `B5`: ni el 30 ni el 60 entran
            # en la fórmula (§1.4). El primer `IF` es el que impide que una
            # fila sin fecha enseñe «VENCIDO hace 46.000 d».
            _f(ws, '{c}{f}'.format(c=col_alerta, f=fila),
               '=IF(${d}{f}="","",'
               'IF(${d}{f}-$B$3<0,"❌ VENCIDO hace "&TEXT($B$3-${d}{f},"0")&" d",'
               'IF(${d}{f}-$B$3<=$B$4,"🔴 "&TEXT(${d}{f}-$B$3,"0")&" d",'
               'IF(${d}{f}-$B$3<=$B$5,"🟡 "&TEXT(${d}{f}-$B$3,"0")&" d",'
               '"🟢 OK"))))'.format(d=col_fecha, f=fila))
            ws['{c}{f}'.format(c=col_alerta, f=fila)].alignment = CENTRO

    nota0 = VENC_R1 + 2
    _nota(ws, nota0,
          '▸ Esta hoja NO se escribe: las cuatro fechas salen de la hoja '
          "'Plantilla' (columnas L, M, N y O) para las 30 filas del "
          'directorio. Si un empleado no aparece, es que le falta la fecha '
          'allí.', ANCHO_VENC, 30)
    _nota(ws, nota0 + 1,
          '▸ Lo único editable son los dos umbrales de arriba (celdas verdes). '
          'Los tres contadores de la derecha cuentan las alertas de las cuatro '
          'columnas a la vez.', ANCHO_VENC, 30)
    _nota(ws, nota0 + 2,
          '▸ El preaviso de fin de contrato temporal de más de un año son 15 '
          'días (art. 49.1.c ET); dejar vencer un temporal sin preaviso o '
          'encadenarlo puede convertirlo en indefinido.', ANCHO_VENC, 30)
    # RT-24 · el aviso de la función volátil, escrito donde se lee.
    _nota(ws, nota0 + 3,
          '▸ «Fecha de hoy» (B3) es =HOY() y Excel la recalcula cada vez que '
          'abres el fichero. En la VISTA PREVIA del móvil o de Google Drive, '
          'que no recalcula, verás la fecha en que descargaste el kit y los '
          'semáforos de ESE día: ábrelo en Excel o LibreOffice antes de '
          'fiarte de los avisos.', ANCHO_VENC, 30)
    merges += ['A{f}:{c}{f}'.format(f=nota0 + i,
                                    c=get_column_letter(ANCHO_VENC))
               for i in range(4)]
    merges.append(_pie(ws, nota0 + 5, ANCHO_VENC))
    _merges(ws, merges)
    cambios.append('07:Vencimientos!A7:I36: 30 empleados × 4 vencimientos '
                   '(contrato, periodo de prueba, carnet de manipulador y '
                   'PRL), leídos del directorio — la v1.1 miraba a 15 de 30 y '
                   'sólo al contrato, y los carnets se tecleaban dos veces '
                   '(DOM-14/TEC-07/COM-16)')
    cambios.append('07:Vencimientos!B4/B5/E3:E5: los dos umbrales del semáforo '
                   'en celda verde y contador de alertas por gravedad — antes '
                   'los 30/60/90 días estaban escritos dentro de las 30 '
                   'fórmulas (§1.4)')


def _n07_instrucciones(wb, cambios):
    """§4 · COM-25/DOM-13 — la sección de protección de datos, reescrita.

    La v1.1 decía tres cosas («proteger con contraseña», «sólo RRHH y
    dirección», «no compartir por email sin cifrar») sin explicar CÓMO, sin
    base jurídica, sin plazos y listando los alérgenos del empleado como un
    dato operativo más.
    """
    ws = wb['Instrucciones']
    motor._limpiar_dv(ws)
    _instrucciones(ws, [
        '07 · Directorio de Plantilla',
        None,
        'Cómo usar esta plantilla:',
        "▸ Registra a cada empleado en una fila de la hoja 'Plantilla'. Caben "
        '30, los mismos que el cuadrante (01) y que el histórico de '
        'evaluaciones (06).',
        '▸ Usa los filtros de Excel para buscar por puesto, convenio o '
        'modalidad de contrato.',
        "▸ La hoja 'Vencimientos' no se escribe: lee las cuatro fechas de "
        'caducidad del directorio y las ordena por gravedad.',
        None,
        'Datos incluidos:',
        '▸ Identificación: nombre, DNI/NIE, fecha de nacimiento y NAF (número '
        'de afiliación a la Seguridad Social).',
        '▸ Relación laboral: puesto, grupo profesional, convenio aplicable, '
        'modalidad de contrato, jornada, horas semanales, alta, fin de '
        'contrato y fin del periodo de prueba.',
        '▸ Formación obligatoria: caducidad del carnet de manipulador de '
        'alimentos y de la formación en PRL.',
        '▸ Contacto: teléfono, email y contacto de emergencia.',
        '▸ Operativos: salario bruto anual, talla de uniforme y taquilla.',
        '▸ NO se piden datos de salud: la columna de alergias e intolerancias '
        'del trabajador que traía la versión 1.1 se ha retirado (ver más '
        'abajo).',
        None,
        'Protección de datos (RGPD):',
        '▸ Base jurídica: el tratamiento de estos datos se ampara en la '
        'ejecución del contrato de trabajo (art. 6.1.b RGPD) y en el '
        'cumplimiento de obligaciones legales laborales, de Seguridad Social '
        'y fiscales (art. 6.1.c). No hace falta pedir consentimiento para '
        'ellos, y pedirlo sería confuso.',
        '▸ Minimización (art. 5.1.c): sólo lo que necesitas para gestionar la '
        'relación laboral. Por eso este directorio ya no pide alergias ni '
        'intolerancias del trabajador: son datos de salud, categoría especial '
        'del art. 9 RGPD, y exigen una base jurídica distinta y reforzada. Si '
        'tu prevención de riesgos necesita alguna, la custodia el servicio de '
        'prevención, no la hoja de RRHH.',
        '▸ Contacto de emergencia: es el dato de un TERCERO. Informa a esa '
        'persona de que la tienes registrada y para qué, y recoge el dato del '
        'empleado sólo con su conocimiento.',
        '▸ Plazos de conservación: el registro diario de jornada se guarda 4 '
        'años (art. 34.9 ET) y la documentación de cotización y de alta, 4 '
        'años (art. 21 LISOS). Pasado el plazo y agotadas las posibles '
        'reclamaciones, se suprime.',
        '▸ Derechos del trabajador: acceso, rectificación, supresión, '
        'limitación, portabilidad y oposición (arts. 15 a 22 RGPD). Ten '
        'previsto a quién se dirige la petición y en cuánto tiempo respondes '
        '(un mes).',
        '▸ Cómo cifrar este libro: Archivo → Información → Proteger libro → '
        'Cifrar con contraseña. Guarda esa contraseña en tu gestor de '
        'contraseñas, no en el nombre del fichero.',
        '▸ Acceso limitado a RRHH y dirección; nunca por email sin cifrar ni '
        'en carpetas compartidas del equipo.',
        None,
        'Nota sobre la protección de las hojas:',
        '▸ Las hojas van protegidas para que nadie borre una fórmula por '
        'error, pero SIN contraseña: son dos cosas distintas. Cifrar el libro '
        '(arriba) es lo que protege los datos personales.',
    ])
    cambios.append('07:Instrucciones!B14+: sección de protección de datos '
                   'reescrita — base jurídica (art. 6.1.b y 6.1.c), '
                   'minimización, plazos (4 años, art. 34.9 ET y art. 21 '
                   'LISOS), contacto de emergencia como dato de tercero, '
                   'derechos de los arts. 15-22 y el CÓMO cifrar el libro '
                   '(COM-25/DOM-13)')


def _n07(wb, cambios):
    _n07_plantilla(wb, cambios)
    _n07_vencimientos(wb, cambios)
    _n07_instrucciones(wb, cambios)


# ==========================================================================
# BONUS-02 · CALCULADORA DE PLANTILLA ÓPTIMA  (SPEC §3)
# ==========================================================================
#: Los 10 tipos de negocio que la propia hoja de referencia ya enumeraba —y que
#: la calculadora de la v1.1 NO admitía: `B8='Tipo negocio (1=casual, 2=fine,
#: 3=fast)'` y `B12='=IF(B8=1,20,IF(B8=2,8,30))'`, así que teclear un 4 devolvía
#: en silencio los ratios de comida rápida (DOM-23/TEC-16).
#:
#: `(tipo, rango cocina, rango sala, rango barra, cocina, sala, barra,
#:   ratio de coste laboral (texto), objetivo, máximo aceptable)`
#:
#: Los tres números son CUBIERTOS POR EMPLEADO Y SERVICIO, no por día: es la
#: unidad con la que se dimensiona un turno y la que usa `03!Previsión por
#: Servicio`, que con estos mismos valores devuelve los mismos 7 FTE (§7-bis.6).
#: El `0` significa «este formato no tiene ese puesto» y la fórmula lo respeta:
#: una dark kitchen no necesita camareros ni barra, y un bar no tiene cocina de
#: carta.
#:
#: Los dos porcentajes son los MISMOS que usa el semáforo del 03 para los seis
#: tipos que comparten los dos ficheros (`grupo_b.TIPOS_NEGOCIO`): el objetivo
#: es el extremo superior del rango de texto que la hoja ya imprimía en la
#: columna H desde la v1.1, y el aceptable, el techo a partir del cual se pide
#: acción correctiva. Dos ficheros del mismo kit no pueden dar veredictos
#: distintos sobre el mismo ratio.
TIPOS_BONUS = [
    ('Restaurante Casual',          '22-28', '20-25', '70-90',
     25, 22, 80,  '28-33%', 0.33, 0.35),
    ('Fine Dining / Alta Cocina',   '10-14', '8-12',  '35-45',
     12, 10, 40,  '35-40%', 0.40, 0.42),
    ('Fast Casual / Comida Rápida', '40-50', '35-45', '90-110',
     45, 40, 100, '25-28%', 0.28, 0.30),
    ('Cafetería / Brunch',          '30-40', '25-35', '50-70',
     35, 30, 60,  '28-32%', 0.32, 0.35),
    ('Bar / Cocktails',             '—',     '25-35', '30-40',
     0,  30, 35,  '25-30%', 0.30, 0.33),
    ('Pizzería',                    '26-34', '22-28', '80-100',
     30, 25, 90,  '26-30%', 0.30, 0.32),
    ('Dark Kitchen / Delivery',     '35-45', '—',     '—',
     40, 0,  0,   '22-28%', 0.28, 0.30),
    ('Hotel (restaurante)',         '18-24', '12-18', '45-55',
     20, 15, 50,  '35-42%', 0.42, 0.44),
    ('Catering / Eventos',          '20-25', '12-18', '50-70',
     22, 15, 60,  '30-35%', 0.35, 0.38),
    ('Heladería / Obrador',         '35-45', '25-35', '—',
     40, 30, 0,   '25-30%', 0.30, 0.32),
]

TABLA_BONUS = "'Ratios por Tipo'!$A$5:$J$14"
ANCHO_CALC = 5
#: Índices del `VLOOKUP` sobre `TABLA_BONUS` (A=1 … J=10).
IDX_COCINA, IDX_SALA, IDX_BARRA = 5, 6, 7
IDX_OBJETIVO, IDX_ACEPTABLE = 9, 10

#: Semáforo del diagnóstico de la comparativa. `VOC_ALERTA` no sirve: aquí lo
#: grave es quedarse CORTO de personal, no pasarse.
VOC_DIMENSION = [('INFRADIMENSIONADA', 'rojo'),
                 ('SOBREDIMENSIONADA', 'ambar'),
                 ('CORRECTAMENTE', 'verde')]


def _cf_propio(ws, rango, vocabulario):
    """Semáforo del grupo. `motor.aplicar_cf` sólo limpia los rangos que
    gobierna ÉL, así que una regla mía escrita dos veces se duplicaría en cada
    pasada y la idempotencia lo cazaría como diferencia de `cf`."""
    motor._limpiar_cf(ws, {rango})
    return motor.semaforo(ws, rango, vocabulario)


def _bonus_ratios(wb, cambios):
    """La tabla de referencia gana las columnas NUMÉRICAS que lee el VLOOKUP."""
    ws = wb['Ratios por Tipo']
    motor._limpiar_dv(ws)
    _merges(ws, [])
    _vaciar(ws)
    _anchos(ws, {'A': 28, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15,
                 'G': 15, 'H': 16, 'I': 15, 'J': 15})
    merges = _titulo(ws, 10, 'Ratios de Personal por Tipo de Negocio')
    _cabecera(ws, 4, [
        'Tipo de Negocio',
        'Cubiertos/cocinero y servicio (rango)',
        'Cubiertos/camarero y servicio (rango)',
        'Cubiertos/barra y servicio (rango)',
        'Cocina: cubiertos (núm.)',
        'Sala: cubiertos (núm.)',
        'Barra: cubiertos (núm.)',
        'Ratio Coste Laboral',
        'Objetivo % (núm.)',
        'Aceptable % (núm.)'])
    for i, fila in enumerate(TIPOS_BONUS):
        r = 5 + i
        for j, valor in enumerate(fila):
            cel = ws.cell(row=r, column=1 + j, value=valor)
            if j >= 1:
                cel.alignment = CENTRO
            # RT-20 · B:D («22-28», «—») llevaban '0' y H («28-33%») '0.0%',
            # heredados del bloque numérico vecino: tres columnas de TEXTO con
            # el formato del tipo contrario. Basta con que alguien teclee un
            # número donde hay un rango para verlo como porcentaje.
            if j in (1, 2, 3, 7):
                cel.number_format = 'General'
        ws.cell(row=r, column=1).alignment = Alignment(vertical='center')
    # 'Ratios por Tipo' está en `motor.SIN_VERDE_AUTO`: el verde de esta hoja
    # se marca a mano y no lo pisa nadie.
    motor.marcar_verde(ws, 'E5:G14')
    motor.marcar_verde(ws, 'I5:J14')

    nota0 = 16
    _nota(ws, nota0,
          '▸ Los tres números son CUBIERTOS POR EMPLEADO Y SERVICIO, no por '
          'día: 80 cubiertos al día en dos servicios son 40 por servicio, y es '
          'con esos 40 con los que se dimensiona el turno.', 10, 30)
    _nota(ws, nota0 + 1,
          '▸ Un 0 significa que ese formato no tiene ese puesto: una dark '
          'kitchen no monta sala y un bar de coctelería no lleva cocina de '
          'carta. La calculadora no pide personal donde el ratio es 0.', 10, 30)
    _nota(ws, nota0 + 2,
          '▸ Las columnas numéricas son las que lee la calculadora por '
          'VLOOKUP; los rangos de al lado son orientativos — ajústalos a tu '
          'carta, a tu servicio y a tu nivel de exigencia.', 10, 30)
    _nota(ws, nota0 + 3,
          '▸ En temporada alta baja el ratio (más personal por cubierto) o '
          'sube el factor de cobertura de la calculadora a 1,20.', 10, 15)
    _nota(ws, nota0 + 4,
          '▸ «Objetivo %» y «Aceptable %» son los mismos umbrales que usa el '
          'semáforo del fichero 03 para los seis tipos que están en los dos: '
          'los dos ficheros no pueden contradecirse sobre el mismo ratio.',
          10, 30)
    _nota(ws, nota0 + 5,
          '▸ El ratio de coste laboral incluye la cotización empresarial, que '
          'en la calculadora es una celda verde (33 % por defecto), no un 30 % '
          'escondido en la fórmula.', 10, 30)
    merges += ['A{f}:J{f}'.format(f=nota0 + i) for i in range(6)]
    merges.append(_pie(ws, nota0 + 7, 10))
    _merges(ws, merges)
    cambios.append('BONUS-02:Ratios por Tipo!E4:J4: tres columnas numéricas de '
                   'cubiertos por empleado y SERVICIO + objetivo y máximo '
                   'aceptable del ratio de coste laboral, que es lo que lee el '
                   'VLOOKUP de la calculadora (DOM-23/TEC-16/DOM-10)')


def _bonus_calculadora(wb, cambios):
    """§3 · DOM-10/DOM-18/DOM-23/TEC-16/TEC-21/COM-13.

    Lo que había, con los datos precargados: 19 empleados, 44.460 €/mes y
    622.440 €/año para un casual de 80 cubiertos —un ratio del 99,7 % frente al
    28-33 % que marcaba su propia hoja de referencia—, «Días apertura/semana»
    como celda MUERTA que ninguna de las 16 fórmulas leía, el ×14 escondido en
    `B29='=B28*14'` y los picos de demanda que la landing vende sin existir.
    """
    ws = wb['Calculadora']
    motor._limpiar_dv(ws)
    _merges(ws, [])
    _vaciar(ws)
    _anchos(ws, {'A': 46, 'B': 16, 'C': 26, 'D': 14, 'E': 18})
    merges = _titulo(ws, ANCHO_CALC, 'Calculadora de Plantilla Óptima')

    _seccion(ws, 4, 'DATOS DE TU NEGOCIO — edita sólo las celdas verdes',
             ANCHO_CALC)
    entradas = [
        (5,  'Cubiertos previstos / día:', 80, motor.FMT_ENT, None),
        (6,  'Servicios al día (comida + cena):', 2, motor.FMT_ENT, None),
        (7,  'Días de apertura / semana:', 6, motor.FMT_ENT,
         'Entra en el cálculo por las horas semanales: abrir 7 días en vez de '
         '6 sube la plantilla necesaria, no sólo el turno. En la versión 1.1 '
         'esta celda no la leía ninguna fórmula.'),
        (8,  'Tipo de negocio:', 'Restaurante Casual', None, None),
        (9,  'Salario medio bruto (€/mes, en las pagas de abajo):', 1500,
         motor.FMT_EUR,
         'Es el BRUTO de convenio de una persona a jornada completa, no su '
         'coste: la cotización empresarial se suma aparte, dos filas más '
         'abajo. Para el coste real de TU equipo usa la hoja «Nóminas» del '
         'fichero 03.'),
        (10, 'Nº de pagas / año:', 14, motor.FMT_ENT, None),
        (11, 'Tipo de SS a cargo de la empresa (%):', 0.33, motor.FMT_PCT1,
         motor.PARAMETROS['ss_empresa'][3]),
        # RD-03 (refutado) · la unidad es HORAS PAGADAS POR PRESENCIA, y una
        # presencia es una persona en UN servicio (B27 = personal/servicio ×
        # servicios). Quien lo lea como «horas de la persona al día» concluirá
        # que faltan la mise en place y el cierre; el rótulo lo dice ahora.
        (12, 'Horas pagadas por presencia y servicio '
             '(el turno completo sale de multiplicarlas por los servicios):',
         4, motor.FMT_ENT,
         'Son las horas que le cuestas a la empresa CADA persona por CADA '
         'servicio que cubre, montaje y cierre incluidos. Con 2 servicios al '
         'día, 4 h por servicio son el turno de 8 h de la hoja «Turnos» del '
         'fichero 01. Súbelas si tu servicio es más largo o si montas y '
         'desmontas fuera del turno.'),
        (13, 'Jornada contratada (h/semana):', 40, motor.FMT_ENT,
         motor.PARAMETROS['jornada_semanal'][3]),
        (14, 'Factor de cobertura (vacaciones, bajas y descansos):', 1.15,
         motor.FMT_DEC2, motor.PARAMETROS['factor_cobertura'][3]),
        (15, 'Ticket medio sin IVA (€):', 25, motor.FMT_EUR,
         'Sólo sirve para contrastar el resultado con tus ventas: sin él la '
         'calculadora puede decirte que contrates a diecinueve personas sin '
         'que nada avise de que te comen el 100 % de la caja.'),
        (16, 'Cubiertos del día PICO:', 120, motor.FMT_ENT,
         'El día fuerte de tu semana (viernes o sábado, casi siempre). De aquí '
         'sale el refuerzo por servicio, que es lo que hay que reforzar ese '
         'día y no toda la semana.'),
    ]
    for fila, rotulo, valor, fmt, nota in entradas:
        _rotulo(ws, fila, rotulo)
        _verde(ws, 'B{}'.format(fila), valor, fmt)
        if nota:
            _nota_celda(ws, 'B{}'.format(fila), rotulo.rstrip(':'), nota)

    _dv(ws, 'B6', ['1', '2', '3'], 'Servicios al día no válido',
        'Uno (sólo comidas o sólo cenas), dos (comida y cena) o tres '
        '(desayuno, comida y cena). Es lo que reparte los cubiertos del día '
        'entre turnos, y el turno es lo que se dimensiona.',
        'Elige 1, 2 o 3.')
    _dv(ws, 'B8', [t[0] for t in TIPOS_BONUS], 'Tipo de negocio no válido',
        'De aquí salen por VLOOKUP los tres ratios de la hoja «Ratios por '
        'Tipo» y los dos umbrales del semáforo del ratio de coste laboral. La '
        'versión 1.1 sólo admitía tres tipos y cualquier otro valor devolvía '
        'en silencio los ratios de comida rápida.',
        'Elige uno de los diez tipos de la hoja «Ratios por Tipo».')
    _dv(ws, 'B10', ['12', '14', '15'], 'Número de pagas no válido',
        'Doce si tu convenio prorratea las extras en la nómina mensual, '
        'catorce si las paga en junio y en diciembre, quince en los convenios '
        'que añaden una de beneficios. El coste anual sale de ESTA celda: la '
        'versión 1.1 multiplicaba por 14 sin decirlo.',
        'Elige 12, 14 o 15.')

    _seccion(ws, 18, 'RATIOS APLICADOS — salen de la hoja «Ratios por Tipo»',
             ANCHO_CALC)
    for fila, rotulo, idx in (
            (19, 'Cubiertos por cocinero y servicio:', IDX_COCINA),
            (20, 'Cubiertos por camarero y servicio:', IDX_SALA),
            (21, 'Cubiertos por persona de barra y servicio:', IDX_BARRA)):
        _rotulo(ws, fila, rotulo)
        _f(ws, 'B{}'.format(fila),
           '=IFERROR(VLOOKUP($B$8,{t},{i},FALSE),0)'
           .format(t=TABLA_BONUS, i=idx), motor.FMT_ENT)
    ws['C19'] = '← 0 = ese formato no tiene ese puesto'
    ws['C19'].font = Font(size=9, italic=True, color='666666')

    _seccion(ws, 23, 'DE CUBIERTOS A PLANTILLA', ANCHO_CALC)
    _rotulo(ws, 24, 'Cubiertos por SERVICIO:')
    _f(ws, 'B24', '=IFERROR(IF(OR($B$5="",$B$6="",$B$6=0),"",'
                  'ROUND($B$5/$B$6,0)),"")', motor.FMT_ENT)
    ws['C24'] = '← es la cifra que dimensiona el turno, no los del día'
    ws['C24'].font = Font(size=9, italic=True, color='666666')

    for i, txt in enumerate(['Cocina', 'Sala', 'Barra', 'Total / servicio']):
        cel = ws.cell(row=25, column=2 + i, value=txt)
        cel.font = Font(bold=True, size=10)
        cel.alignment = CENTRO
    _rotulo(ws, 26, 'Personal por SERVICIO:')
    for col, ratio in (('B', '$B$19'), ('C', '$B$20'), ('D', '$B$21')):
        _f(ws, '{}26'.format(col),
           '=IFERROR(IF(OR($B$24="",{r}="",{r}=0),0,'
           'CEILING($B$24/{r},1)),0)'.format(r=ratio), motor.FMT_ENT)
    _f(ws, 'E26', '=IF($B$24="","",$B$26+$C$26+$D$26)', motor.FMT_ENT)
    _rotulo(ws, 27, 'Presencias al día (personas × servicios):')
    _f(ws, 'B27', '=IF(OR($E$26="",$B$6=""),"",$E$26*$B$6)', motor.FMT_ENT)
    _rotulo(ws, 28, 'Horas de plantilla necesarias / semana:')
    _f(ws, 'B28', '=IF(OR($B$27="",$B$12="",$B$7=""),"",'
                  '$B$27*$B$12*$B$7)', motor.FMT_ENT)
    _rotulo(ws, 29, 'Plantilla necesaria (FTE):', negrita=True)
    _f(ws, 'B29', '=IFERROR(IF(OR($B$28="",$B$13="",$B$13=0),"",'
                  'CEILING($B$28/$B$13*$B$14,1)),"")', motor.FMT_ENT)
    ws['B29'].font = NEGRITA

    _seccion(ws, 31, 'REFUERZO EN DÍA PICO', ANCHO_CALC)
    _rotulo(ws, 32, 'Cubiertos por servicio en el día pico:')
    _f(ws, 'B32', '=IFERROR(IF(OR($B$16="",$B$6="",$B$6=0),"",'
                  'ROUND($B$16/$B$6,0)),"")', motor.FMT_ENT)
    for i, txt in enumerate(['Cocina', 'Sala', 'Barra', 'Total / servicio']):
        cel = ws.cell(row=33, column=2 + i, value=txt)
        cel.font = Font(size=9, color='666666')
        cel.alignment = CENTRO
    _rotulo(ws, 34, 'Personal por servicio ese día:')
    for col, ratio in (('B', '$B$19'), ('C', '$B$20'), ('D', '$B$21')):
        _f(ws, '{}34'.format(col),
           '=IFERROR(IF(OR($B$32="",{r}="",{r}=0),0,'
           'CEILING($B$32/{r},1)),0)'.format(r=ratio), motor.FMT_ENT)
    _f(ws, 'E34', '=IF($B$32="","",$B$34+$C$34+$D$34)', motor.FMT_ENT)
    _rotulo(ws, 35, 'REFUERZO sobre un día normal (personas por servicio):',
            negrita=True)
    _f(ws, 'B35', '=IF(OR($E$34="",$E$26=""),"",MAX(0,$E$34-$E$26))',
       motor.FMT_ENT)
    ws['B35'].font = NEGRITA
    ws['C35'] = '← es un refuerzo de ESE día, no plantilla fija'
    ws['C35'].font = Font(size=9, italic=True, color='666666')

    _seccion(ws, 37, 'COSTE LABORAL Y CONTRASTE CON LAS VENTAS', ANCHO_CALC)
    _rotulo(ws, 38, 'Coste laboral estimado del equipo (€/mes):')
    _f(ws, 'B38', '=IFERROR(IF(OR($B$29="",$B$9="",$B$10=""),"",'
                  'ROUND($B$29*$B$9*$B$10/12*(1+$B$11),2)),"")', motor.FMT_EUR)
    _rotulo(ws, 39, 'Coste laboral estimado (€/año):')
    _f(ws, 'B39', '=IF($B$38="","",ROUND($B$38*12,2))', motor.FMT_EUR)
    ws['C39'] = ('← doce meses del coste mensual, que ya lleva dentro las '
                 'pagas extra')
    ws['C39'].font = Font(size=9, italic=True, color='666666')
    # 52/12 no es un parámetro del negocio: son las semanas que tiene un mes.
    _rotulo(ws, 40, 'Ventas estimadas del mes (€):')
    _f(ws, 'B40', '=IFERROR(IF(OR($B$5="",$B$15="",$B$7=""),"",'
                  'ROUND($B$5*$B$15*$B$7*52/12,2)),"")', motor.FMT_EUR)
    _rotulo(ws, 41, 'RATIO DE COSTE LABORAL:', negrita=True)
    _f(ws, 'B41', '=IFERROR(IF(OR($B$38="",$B$40="",$B$40=0),"",'
                  'ROUND($B$38/$B$40,4)),"")', motor.FMT_PCT1)
    ws['B41'].font = NEGRITA
    _rotulo(ws, 41, 'Ratio objetivo de tu tipo:', col=3)
    _f(ws, 'D41', '=IFERROR(VLOOKUP($B$8,{t},{i},FALSE),0.33)'
       .format(t=TABLA_BONUS, i=IDX_OBJETIVO), motor.FMT_PCT1)
    _rotulo(ws, 42, 'Máximo aceptable de tu tipo:', col=3)
    _f(ws, 'D42', '=IFERROR(VLOOKUP($B$8,{t},{i},FALSE),0.35)'
       .format(t=TABLA_BONUS, i=IDX_ACEPTABLE), motor.FMT_PCT1)
    _rotulo(ws, 43, 'Semáforo:', negrita=True)
    _f(ws, 'B43',
       '=IF($B$41="","— introduce el ticket medio y elige tu tipo de negocio",'
       'IF($B$41<$D$41,"🟢 EXCELENTE (por debajo de tu objetivo)",'
       'IF($B$41<=$D$42,"🟡 VIGILAR (en el límite)",'
       '"🔴 ACCIÓN CORRECTIVA")))')
    ws['B43'].alignment = Alignment(vertical='center')
    merges.append('B43:E43')
    _cf_propio(ws, 'B43:B43', motor.VOC_RATIO)

    _seccion(ws, 45, 'COMPARATIVA CON TU PLANTILLA ACTUAL', ANCHO_CALC)
    _rotulo(ws, 46, 'Personal actual (personas a jornada completa):')
    _verde(ws, 'B46', None, motor.FMT_ENT)
    _rotulo(ws, 47, 'Diferencia:')
    _f(ws, 'B47', '=IF(OR($B$46="",$B$29=""),"",$B$46-$B$29)', motor.FMT_ENT)
    _rotulo(ws, 48, 'Diagnóstico:')
    _f(ws, 'B48',
       '=IF($B$47="","",'
       'IF($B$47>1,"⚠ SOBREDIMENSIONADA (+"&TEXT($B$47,"0")&")",'
       'IF($B$47<-1,"🔴 INFRADIMENSIONADA ("&TEXT($B$47,"0")&")",'
       '"✓ DIMENSIONADA CORRECTAMENTE")))')
    ws['B48'].alignment = Alignment(vertical='center')
    merges.append('B48:E48')
    _cf_propio(ws, 'B48:B48', VOC_DIMENSION)

    nota0 = 50
    _nota(ws, nota0,
          '▸ FTE = personas a jornada completa, no personas por turno. Y una '
          'PRESENCIA no es una persona: es una persona en UN servicio. Con '
          'los datos de ejemplo son 5 personas por servicio y 2 servicios al '
          'día = 10 presencias, que son esas MISMAS 5 personas haciendo un '
          'turno de 8 h (4 h pagadas por servicio × 2): 5 × 8 h × 6 días = '
          '240 h a la semana = siete contratos de 40 h con el 15 % de '
          'cobertura de libranzas, vacaciones y bajas ya dentro. Cuadra con '
          "la hoja 'Turnos' del 01 (M, T y N son de 8 h; el partido, 9).",
          ANCHO_CALC, 46)
    # RC-12 · la nota anterior decía que una discrepancia con el 03 sólo
    # podía venir de «un parámetro cambiado en uno y no en el otro». Es falso
    # como diagnóstico: 03!'Previsión por Servicio' NO tiene selector de tipo
    # de negocio —sus tres ratios son constantes de Restaurante Casual—, así
    # que en cuanto aquí se elige otro tipo los dos ficheros divergen POR
    # DISEÑO y el cliente se ponía a buscar un error que no había cometido.
    _nota(ws, nota0 + 1,
          '▸ La hoja «Previsión por Servicio» del fichero 03 hace este mismo '
          'cálculo y con los valores de fábrica devuelve el mismo número. Si '
          'cambias aquí el TIPO DE NEGOCIO, esa hoja no se entera: sus tres '
          'ratios son constantes de Restaurante Casual. Copia entonces B19, '
          'B20 y B21 de esta calculadora a B7, B8 y B9 de «Previsión por '
          'Servicio» y volverán a coincidir.', ANCHO_CALC, 42)
    _nota(ws, nota0 + 2,
          '▸ El refuerzo del día pico es personal EXTRA de ese día (horas '
          'complementarias, un fijo-discontinuo o un contrato por '
          'circunstancias de la producción), no plantilla fija: por eso no se '
          'suma al FTE.', ANCHO_CALC, 30)
    _nota(ws, nota0 + 3,
          '▸ El coste es una estimación de convenio, no tu nómina: parte de un '
          'salario medio y de un tipo de cotización únicos para toda la '
          'plantilla. El coste real, contrato a contrato, sale de la hoja '
          '«Nóminas» del fichero 03.', ANCHO_CALC, 30)
    merges += ['A{f}:{c}{f}'.format(f=nota0 + i,
                                    c=get_column_letter(ANCHO_CALC))
               for i in range(4)]
    merges.append(_pie(ws, nota0 + 5, ANCHO_CALC))
    _merges(ws, merges)
    cambios.append('BONUS-02:Calculadora!B24:B29: un solo modelo de '
                   'dimensionamiento (cubiertos por SERVICIO → 5 personas por '
                   'servicio → 10 presencias → 240 h/semana → 7 FTE), el mismo '
                   'que el 03 — la v1.1 daba 19 empleados y un ratio del '
                   '99,7 % (DOM-10/COM-09)')
    cambios.append('BONUS-02:Calculadora!B7/B28: «Días de apertura / semana» '
                   'entra por fin en el cálculo, a través de las horas '
                   'semanales — era una celda muerta que ninguna de las 16 '
                   'fórmulas leía (DOM-18/COM-13)')
    cambios.append('BONUS-02:Calculadora!B8/B19:B21: el tipo de negocio pasa a '
                   'DV de lista con los 10 nombres de la tabla y VLOOKUP, con '
                   '0 = puesto inexistente — antes eran tres IF anidados con '
                   'cajón de sastre a fast casual (DOM-23/TEC-16)')
    cambios.append('BONUS-02:Calculadora!B10/B39: el ×14 escondido pasa a la '
                   'celda de pagas y el coste anual son doce meses del coste '
                   'mensual, que ya las lleva dentro (TEC-21)')
    cambios.append('BONUS-02:Calculadora!B32:B35: refuerzo del día PICO '
                   '(personal del día pico − personal del día normal), que es '
                   'lo que la landing vende como «picos de demanda» y no '
                   'existía (COM-13)')
    cambios.append('BONUS-02:Calculadora!B40:B42: ventas estimadas, ratio de '
                   'coste laboral y semáforo contra el objetivo de la tabla — '
                   'el contraste que impide volver a entregar un resultado del '
                   '99,7 % sin un solo aviso (DOM-10)')


def _bonus_instrucciones(wb, cambios):
    ws = wb['Instrucciones']
    motor._limpiar_dv(ws)
    _instrucciones(ws, [
        'BONUS · Calculadora de Plantilla Óptima',
        None,
        'Cómo usar esta plantilla:',
        "▸ Introduce los datos de tu negocio en las celdas verdes de la hoja "
        "'Calculadora'. Todo lo demás son fórmulas.",
        '▸ El tipo de negocio es un desplegable con los diez formatos de la '
        "hoja 'Ratios por Tipo': de él salen los tres ratios y los dos "
        'umbrales del semáforo.',
        '▸ Compara el resultado con tu plantilla actual en el bloque del '
        'final para detectar sobre o infradimensionamiento.',
        None,
        'La cadena de cálculo, en una línea:',
        '▸ Cubiertos al día ÷ servicios = cubiertos por SERVICIO → ÷ ratio de '
        'cada puesto = personal por servicio → × servicios = presencias al '
        'día → × horas por servicio × días de apertura = horas a la semana → '
        '÷ jornada × factor de cobertura = FTE.',
        '▸ Con los datos de ejemplo (80 cubiertos, 2 servicios, casual): 40 '
        'cubiertos por servicio, 5 personas por servicio, 10 presencias al '
        'día, 240 h a la semana y 7 personas a jornada completa.',
        '▸ FTE son contratos a jornada completa, no personas por turno. En '
        'sala, a la vez, hay 2.',
        None,
        'Por qué hay que meter el ticket medio:',
        '▸ Sin él, la calculadora puede decirte que contrates a diecinueve '
        'personas sin que nada avise de que se comen toda la caja. Con él, la '
        'hoja estima tus ventas y pone el ratio de coste laboral al lado del '
        'objetivo de tu tipo de negocio.',
        '▸ Con los datos de ejemplo: 52.000 € de ventas al mes, 16.292,50 € '
        'de coste y un ratio del 31,3 %, dentro del 28-33 % de un casual.',
        None,
        'El día pico:',
        '▸ «Cubiertos del día PICO» es tu día fuerte (viernes o sábado, casi '
        'siempre). La hoja calcula el personal de ESE día y te da el refuerzo '
        'sobre un día normal: es lo que hay que reforzar ese día, no toda la '
        'semana, y no se suma a la plantilla fija.',
        None,
        'Lo que esta hoja NO es:',
        '▸ No es tu coste de nómina: parte de un salario medio y de un tipo de '
        'cotización únicos para todo el equipo. El coste real, contrato a '
        'contrato, sale de la hoja «Nóminas» del fichero 03 — que es otro '
        'libro, no una pestaña de éste.',
        '▸ Los ratios son orientativos y editables: si tu carta es más '
        'laboriosa o tu servicio más exigente, baja el número de cubiertos por '
        "cocinero o por camarero en las celdas verdes de 'Ratios por Tipo' y "
        'la calculadora lo recoge sola.',
    ])
    cambios.append('BONUS-02:Instrucciones: la cadena de cálculo explicada, el '
                   'porqué del ticket medio, el día pico y el aviso de que '
                   'esto no es la nómina (COM-13/TEC-21)')


def _bonus(wb, cambios):
    _bonus_ratios(wb, cambios)
    _bonus_calculadora(wb, cambios)
    _bonus_instrucciones(wb, cambios)


# ==========================================================================
# Contrato con main.py
# ==========================================================================
def pre(wb, fname, cambios):
    """Todo el trabajo del grupo, ANTES de `motor.aplicar`."""
    if fname.startswith('06-'):
        _n06(wb, cambios)
    elif fname.startswith('07-'):
        _n07(wb, cambios)
    elif fname.startswith('BONUS-02'):
        _bonus(wb, cambios)
    return cambios


# ==========================================================================
# Demostraciones con pycel (SPEC §5) — se cambian ENTRADAS y se comprueba la
# DIRECCIÓN del resultado, que es lo único que demuestra que la cadena existe.
# ==========================================================================
EPOCA = datetime.datetime(1899, 12, 30)


def _serie(y, m, d):
    """Fecha → número de serie de Excel. pycel opera con el serial, no con el
    `datetime`: pasarle un objeto fecha en `set_value` deja la resta de
    `$B7-$B$3` en `#VALUE!`."""
    return (datetime.datetime(y, m, d) - EPOCA).days


def _hoy():
    return (datetime.datetime.now().replace(hour=0, minute=0, second=0,
                                            microsecond=0) - EPOCA).days


def _xl(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                              # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)


def _sv(xl, ref, valor):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            xl.evaluate(ref)
            xl.set_value(ref, valor)
            return True
        except Exception:                                   # noqa: BLE001
            return False


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _calentar(xl, refs):
    """⚠ pycel sólo propaga un `set_value` a las celdas que YA están en su
    grafo de dependencias: hay que evaluar primero las SALIDAS para que la
    cadena entera entre en el grafo y la invalidación funcione."""
    return dict((r, _ev(xl, r)) for r in refs)


def _d06_ficha(carpeta):
    """§4 · la ficha en blanco NO enseña `#¡DIV/0!` y el `N/A` sale de la media.

    La v1.1 entregaba `C22` y `C23` con el error CACHEADO en un documento que
    se imprime y se firma (DOM-02/TEC-03/COM-04), y su DV `whole 1..5` no
    admitía «no aplica»: la única forma de no puntuar una competencia era
    dejarla en blanco, indistinguible de un olvido (DOM-30).
    """
    p = os.path.join(carpeta, F06)
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    refs = ["'Ficha Evaluación'!C22", "'Ficha Evaluación'!C23",
            "'Ficha Evaluación'!C24"]
    vacia = _calentar(xl, refs)
    notas = [5, 5, 4, 4, 5, 4, 'N/A', 4, 3, 4]      # el mismo caso del ejemplo
    for i, nota in enumerate(notas):
        _sv(xl, "'Ficha Evaluación'!C{}".format(FICHA_R0 + i), nota)
    con_na = dict((r, _ev(xl, r)) for r in refs)
    # la novena competencia sube de 3 a 5: la media TIENE que subir
    _sv(xl, "'Ficha Evaluación'!C20", 5)
    subida = _ev(xl, refs[0])
    # RD-24 · el caso que abría el hallazgo: una sola competencia puntuada.
    solo = _xl(p)
    _calentar(solo, refs)
    _sv(solo, "'Ficha Evaluación'!C{}".format(FICHA_R0), 5)
    una_sola = dict((r.split('!')[1], str(_ev(solo, r))) for r in refs)
    ejemplo = _xl(p)
    ej = dict((r.replace('Ficha Evaluación', 'Ficha (ejemplo relleno)'),
               _ev(ejemplo, r.replace('Ficha Evaluación',
                                      'Ficha (ejemplo relleno)')))
              for r in refs)
    media = con_na["'Ficha Evaluación'!C22"]
    # RD-24 · con la ficha en blanco el NIVEL ya no calla: reclama que se
    # puntúen al menos MIN_COMPETENCIAS. Callar era lo que dejaba emitir
    # «⭐ EXCELENTE» con UNA competencia rellena, en el documento que se firma.
    ok = (vacia["'Ficha Evaluación'!C22"] in ('', None)
          and isinstance(vacia["'Ficha Evaluación'!C23"], str)
          and 'puntúa al menos' in vacia["'Ficha Evaluación'!C23"]
          and vacia["'Ficha Evaluación'!C24"] == '0 de 10'
          and _num(media) and abs(media - 4.22) < 0.005
          and con_na["'Ficha Evaluación'!C24"] == '9 de 10'
          and con_na["'Ficha Evaluación'!C23"] == '✓ BUENO'
          and _num(subida) and subida > media
          and 'puntúa al menos' in una_sola['C23']
          and una_sola['C24'] == '1 de 10')
    return {'ref': '06-evaluacion-desempeno.xlsx:Ficha Evaluación:C22/C23/C24',
            'formula_media': motor.guarda_media('$C$12:$C$21'),
            'ficha_en_blanco': dict((k.split('!')[1], str(v))
                                    for k, v in vacia.items()),
            'nueve_notas_y_una_NA': dict((k.split('!')[1], str(v))
                                         for k, v in con_na.items()),
            'media_al_subir_C20_de_3_a_5': subida,
            'hoja_de_ejemplo': dict((k.split('!')[1], str(v))
                                    for k, v in ej.items()),
            'ok': ok,
            'una_sola_competencia_puntuada': una_sola,
            'nota': 'con 9 competencias puntuadas y una N/A la media son '
                    '38/9 = 4,22 (no 38/10 = 3,80): el N/A no penaliza, y el '
                    'contador dice sobre cuántas se ha calculado. Y con UNA '
                    'sola competencia puntuada el nivel NO se emite (RD-24): '
                    'antes C12=5 y las otras nueve en blanco daban '
                    '«⭐ EXCELENTE»'}


def _d06_tendencia(carpeta):
    """§4 · la tendencia compara los DOS ÚLTIMOS trimestres informados.

    La v1.1 miraba sólo Q4 contra Q3 (`=IF(AND(E5<>"",D5<>""),…)`), así que la
    columna estaba en blanco nueve meses al año (TEC-27/COM-21).
    """
    p = os.path.join(carpeta, F06)
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    _calentar(xl, ["'Histórico'!F5", "'Histórico'!G5"])
    pasos = []
    for etiqueta, valores, esperado in (
            ('sólo Q1 = 3,0 (un dato: no hay con qué comparar)',
             [3, None, None, None], ''),
            ('Q1 = 3,0 y Q2 = 4,0 (la v1.1 seguía en blanco)',
             [3, 4, None, None], '↑ Mejora'),
            ('Q1, Q2 y Q3 = 2,0', [3, 4, 2, None], '↓ Baja'),
            ('Q1, Q2, Q3 y Q4 = 2,0', [3, 4, 2, 2], '→ Estable')):
        for i, v in enumerate(valores):
            _sv(xl, "'Histórico'!{}5".format(get_column_letter(2 + i)),
                v if v is not None else None)
        obtenido = _ev(xl, "'Histórico'!G5")
        media = _ev(xl, "'Histórico'!F5")
        pasos.append({'caso': etiqueta, 'esperado': esperado,
                      'obtenido': obtenido, 'media_anual': media,
                      'ok': obtenido == esperado})
    return {'ref': '06-evaluacion-desempeno.xlsx:Histórico:G5',
            'pruebas': pasos, 'ok': all(x['ok'] for x in pasos),
            'nota': 'la comparación es INDEX sobre COUNT: el último trimestre '
                    'informado contra el anterior, sin `LOOKUP(9^9;…)`'}


def _d07_vencimientos(carpeta):
    """§4 · los CUATRO vencimientos y los 30 empleados.

    La v1.1 leía `Plantilla!A5:A19` —15 de 30— y sólo el fin de contrato: del
    empleado 16 en adelante nadie avisaba (DOM-14/TEC-07). Se prueba sobre el
    EMPLEADO 30 (`Plantilla` fila 34 → `Vencimientos` fila 36), que es
    justamente el que la v1.1 no miraba, y sobre el carnet de manipulador, que
    antes se tecleaba a mano en otra tabla.
    """
    p = os.path.join(carpeta, F07)
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    hoy = _hoy()
    salidas = ["'Vencimientos'!C36", "'Vencimientos'!G36",
               "'Vencimientos'!E3", "'Vencimientos'!E4", "'Vencimientos'!E5"]
    vacio = _calentar(xl, salidas)
    pruebas = []
    for etiqueta, dias, empieza in (
            ('contrato del empleado 30 vencido hace 10 días', -10, '❌'),
            ('contrato del empleado 30 a 15 días (umbral rojo = 30)', 15,
             '🔴'),
            ('contrato del empleado 30 a 45 días (umbral ámbar = 60)', 45,
             '🟡'),
            ('contrato del empleado 30 a 200 días', 200, '🟢')):
        _sv(xl, "'Plantilla'!L34", hoy + dias)
        v = _ev(xl, "'Vencimientos'!C36")
        pruebas.append({'caso': etiqueta, 'empieza_por': empieza,
                        'obtenido': v,
                        'ok': isinstance(v, str) and v.startswith(empieza)})
    # el carnet de manipulador, que la v1.1 no leía del directorio
    _sv(xl, "'Plantilla'!N34", hoy + 5)
    carnet = _ev(xl, "'Vencimientos'!G36")
    pruebas.append({'caso': 'carnet de manipulador del empleado 30 a 5 días',
                    'empieza_por': '🔴', 'obtenido': carnet,
                    'ok': isinstance(carnet, str) and carnet.startswith('🔴')})
    contadores = dict((r.split('!')[1], _ev(xl, r)) for r in salidas[2:])
    pruebas.append({'caso': 'contador de alertas en ROJO de la cabecera',
                    'empieza_por': '1 — al acabar la batería el contrato está '
                                   'a 200 días (verde) y sólo queda el carnet '
                                   'de manipulador a 5 días',
                    'obtenido': contadores.get('E4'),
                    'ok': contadores.get('E4') == 1})
    return {'ref': '07-directorio-plantilla.xlsx:Vencimientos:C36/G36/E3:E5',
            'fila_probada': 'Vencimientos!36 ← Plantilla!34 (empleado 30 de '
                            '30; la v1.1 llegaba al 15)',
            'recien_descargado': dict((k.split('!')[1], str(v))
                                      for k, v in vacio.items()),
            'pruebas': pruebas,
            'contadores_finales': contadores,
            'ok': all(x['ok'] for x in pruebas),
            'nota': 'los dos umbrales (30 y 60 días) son celdas verdes B4 y '
                    'B5, no literales dentro de las 120 fórmulas de alerta'}


def _d07_menor_edad(carpeta):
    """§4 · DOM-24 — la fecha de nacimiento entró para ESTO.

    Sin ella nadie puede saber que a esa persona no se le puede poner un turno
    de noche ni una hora extra (art. 6 ET).
    """
    p = os.path.join(carpeta, F07)
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    ref = "'Plantilla'!{}5".format(COL_AVISO)
    vacio = _calentar(xl, [ref])[ref]
    hoy = datetime.datetime.now()
    pruebas = []
    for etiqueta, anos, avisa in (('nacido hace 17 años', 17, True),
                                  ('nacido hace 19 años', 19, False),
                                  ('nacido hace 34 años', 34, False)):
        _sv(xl, "'Plantilla'!C5",
            _serie(hoy.year - anos, hoy.month, min(hoy.day, 28)))
        v = _ev(xl, ref)
        salta = isinstance(v, str) and v.startswith('⚠')
        pruebas.append({'caso': etiqueta, 'esperaba_aviso': avisa,
                        'obtenido': v, 'ok': salta == avisa})

    # RD-15/RT-07 · el borde que la constante 6570 (= 18 × 365) se comía. Entre
    # 2008 y 2026 hay cinco bisiestos, así que 18 años reales son 6.574-6.575
    # días: con la fórmula vieja, quien llevaba 6.570 días vivo YA no recibía
    # el aviso pese a seguir siendo menor. Se prueban los tres días del borde
    # por FECHA, que es lo único que no depende de bisiestos.
    for etiqueta, dias_desplaza, avisa in (
            ('cumple 18 MAÑANA (aún menor)', 1, True),
            ('cumple 18 HOY (ya mayor)', 0, False),
            ('cumplió 18 AYER', -1, False),
            ('hoy − 6.570 días (18 × 365: aún menor por los bisiestos)',
             None, True)):
        if dias_desplaza is None:
            nacimiento = hoy - datetime.timedelta(days=6570)
        else:
            cumple = hoy + datetime.timedelta(days=dias_desplaza)
            nacimiento = datetime.datetime(cumple.year - 18, cumple.month,
                                           min(cumple.day, 28))
        _sv(xl, "'Plantilla'!C5",
            _serie(nacimiento.year, nacimiento.month, nacimiento.day))
        v = _ev(xl, ref)
        salta = isinstance(v, str) and v.startswith('⚠')
        pruebas.append({'caso': etiqueta,
                        'nacimiento': nacimiento.strftime('%d/%m/%Y'),
                        'esperaba_aviso': avisa, 'obtenido': v,
                        'ok': salta == avisa})
    return {'ref': '07-directorio-plantilla.xlsx:Plantilla:{}5'
                   .format(COL_AVISO),
            'sin_fecha_de_nacimiento': str(vacio),
            'pruebas': pruebas, 'ok': all(x['ok'] for x in pruebas),
            'nota': 'la columna «Aviso» es la única calculada del directorio; '
                    'con la casilla vacía se queda EN BLANCO, no en 0. La '
                    'mayoría de edad se mide con DATE(YEAR+18,MONTH,DAY) y no '
                    'contando 6.570 días, que apagaba el aviso hasta cinco '
                    'días antes del cumpleaños (RD-15/RT-07)'}


def _d_bonus_fte(carpeta):
    """§3 · DOM-10/DOM-18/DOM-23 — 7 FTE, ratio 31,3 % y las entradas VIVAS.

    La v1.1 daba 19 empleados y 622.440 €/año para el mismo casual de 80
    cubiertos —un ratio del 99,7 %—, «Días apertura/semana» no la leía ninguna
    fórmula y el tipo de negocio caía en fast casual con cualquier valor
    distinto de 1 o 2.
    """
    p = os.path.join(carpeta, FB2)
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    salidas = ["'Calculadora'!B24", "'Calculadora'!E26", "'Calculadora'!B27",
               "'Calculadora'!B28", "'Calculadora'!B29", "'Calculadora'!B38",
               "'Calculadora'!B39", "'Calculadora'!B40", "'Calculadora'!B41",
               "'Calculadora'!B43", "'Calculadora'!B19", "'Calculadora'!C26"]
    base = _calentar(xl, salidas)
    fte0 = base["'Calculadora'!B29"]
    ratio0 = base["'Calculadora'!B41"]
    pruebas = [{'caso': 'valores por defecto (80 cubiertos, 2 servicios, '
                        'casual, 6 días, 1.500 € en 14 pagas)',
                'esperado': '7 FTE · 16.292,50 €/mes · 195.510 €/año · '
                            '52.000 € de ventas · ratio 31,3 % · 🟢',
                'obtenido': dict((k.split('!')[1], v)
                                 for k, v in base.items()),
                'ok': (fte0 == 7
                       and abs(base["'Calculadora'!B38"] - 16292.50) < 0.01
                       and abs(base["'Calculadora'!B39"] - 195510.0) < 0.01
                       and abs(base["'Calculadora'!B40"] - 52000.0) < 0.01
                       and abs(ratio0 - 0.3133) < 0.0001
                       and isinstance(base["'Calculadora'!B43"], str)
                       and 'EXCELENTE' in base["'Calculadora'!B43"])}]

    # DOM-18: los días de apertura ENTRAN en el cálculo — la celda muerta
    _sv(xl, "'Calculadora'!B7", 7)
    fte7 = _ev(xl, "'Calculadora'!B29")
    horas7 = _ev(xl, "'Calculadora'!B28")
    pruebas.append({'caso': 'abrir 7 días en vez de 6 (celda B7, muerta en la '
                            'v1.1)',
                    'esperado': 'más horas semanales y más FTE',
                    'obtenido': {'B28': horas7, 'B29': fte7},
                    'ok': _num(fte7) and fte7 > fte0 and horas7 == 280})
    _sv(xl, "'Calculadora'!B7", 4)
    fte4 = _ev(xl, "'Calculadora'!B29")
    pruebas.append({'caso': 'abrir 4 días', 'esperado': 'menos FTE que con 6',
                    'obtenido': {'B29': fte4},
                    'ok': _num(fte4) and fte4 < fte0})
    _sv(xl, "'Calculadora'!B7", 6)

    # DOM-23/TEC-16: los 10 tipos por VLOOKUP, y el 0 = puesto inexistente
    _sv(xl, "'Calculadora'!B8", 'Dark Kitchen / Delivery')
    dark = dict((c, _ev(xl, "'Calculadora'!{}".format(c)))
                for c in ('B19', 'B20', 'B21', 'B26', 'C26', 'D26', 'E26',
                          'B29', 'D41'))
    pruebas.append({'caso': 'tipo de negocio = Dark Kitchen / Delivery',
                    'esperado': 'sala y barra a 0 personas (no hay sala) y el '
                                'objetivo del semáforo baja a 28 %',
                    'obtenido': dark,
                    'ok': (dark['C26'] == 0 and dark['D26'] == 0
                           and dark['B26'] > 0
                           and abs(dark['D41'] - 0.28) < 0.0001)})
    _sv(xl, "'Calculadora'!B8", 'Fine Dining / Alta Cocina')
    fine = dict((c, _ev(xl, "'Calculadora'!{}".format(c)))
                for c in ('B19', 'E26', 'B29', 'B41', 'D41', 'B43'))
    pruebas.append({'caso': 'tipo de negocio = Fine Dining / Alta Cocina',
                    'esperado': 'más plantilla que un casual y objetivo 40 %',
                    'obtenido': fine,
                    'ok': (_num(fine['B29']) and fine['B29'] > fte0
                           and abs(fine['D41'] - 0.40) < 0.0001)})
    _sv(xl, "'Calculadora'!B8", 'Restaurante Casual')

    # DOM-10: el contraste con las ventas, que es lo que impide volver a
    # entregar un 99,7 % sin un solo aviso
    _sv(xl, "'Calculadora'!B15", 12)
    caro = dict((c, _ev(xl, "'Calculadora'!{}".format(c)))
                for c in ('B40', 'B41', 'B43'))
    pruebas.append({'caso': 'ticket medio de 12 € en vez de 25 €',
                    'esperado': 'el ratio se dispara y el semáforo pide '
                                'acción correctiva',
                    'obtenido': caro,
                    'ok': (_num(caro['B41']) and caro['B41'] > ratio0
                           and isinstance(caro['B43'], str)
                           and 'ACCIÓN CORRECTIVA' in caro['B43'])})
    _sv(xl, "'Calculadora'!B15", 25)

    # TEC-21: el ×14 escondido pasa a la celda de pagas
    _sv(xl, "'Calculadora'!B10", 12)
    doce = dict((c, _ev(xl, "'Calculadora'!{}".format(c)))
                for c in ('B38', 'B39'))
    pruebas.append({'caso': 'convenio que prorratea (12 pagas en vez de 14)',
                    'esperado': 'el coste baja; en la v1.1 el 14 estaba '
                                'escrito dentro de `=B28*14`',
                    'obtenido': doce,
                    'ok': (_num(doce['B38'])
                           and doce['B38'] < base["'Calculadora'!B38"])})
    _sv(xl, "'Calculadora'!B10", 14)
    return {'ref': 'BONUS-02-calculadora-plantilla-optima.xlsx:Calculadora:'
                   'B29 (FTE) y B41 (ratio)',
            'pruebas': pruebas, 'ok': all(x['ok'] for x in pruebas),
            'nota': 'el mismo modelo que `03!Previsión por Servicio`: '
                    'cubiertos por SERVICIO → personal por servicio → '
                    'presencias/día → horas/semana → FTE (`main.demo_fte` '
                    'compara los dos)'}


def _d_bonus_pico(carpeta):
    """§3 · COM-13 — «picos de demanda» deja de ser una promesa de la landing.

    En la v1.1 no había ninguna celda de día pico: la calculadora daba una sola
    cifra para toda la semana.
    """
    p = os.path.join(carpeta, FB2)
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    salidas = ["'Calculadora'!B32", "'Calculadora'!E34", "'Calculadora'!B35",
               "'Calculadora'!E26", "'Calculadora'!B29"]
    base = _calentar(xl, salidas)
    pruebas = [{'caso': 'día pico de 120 cubiertos contra 80 de un día normal',
                'esperado': '60 por servicio, 7 personas por servicio y 2 de '
                            'refuerzo',
                'obtenido': dict((k.split('!')[1], v)
                                 for k, v in base.items()),
                'ok': (base["'Calculadora'!B32"] == 60
                       and base["'Calculadora'!E34"] == 7
                       and base["'Calculadora'!B35"] == 2)}]
    _sv(xl, "'Calculadora'!B16", 80)
    igual = _ev(xl, "'Calculadora'!B35")
    pruebas.append({'caso': 'día pico igual que un día normal (80)',
                    'esperado': '0 de refuerzo', 'obtenido': igual,
                    'ok': igual == 0})
    _sv(xl, "'Calculadora'!B16", 200)
    fuerte = dict((c, _ev(xl, "'Calculadora'!{}".format(c)))
                  for c in ('B32', 'E34', 'B35', 'B29'))
    pruebas.append({'caso': 'día pico de 200 cubiertos',
                    'esperado': 'más refuerzo, y el FTE NO cambia (es '
                                'personal extra de ese día)',
                    'obtenido': fuerte,
                    'ok': (_num(fuerte['B35']) and fuerte['B35'] > 2
                           and fuerte['B29'] == base["'Calculadora'!B29"])})
    _sv(xl, "'Calculadora'!B16", 120)
    return {'ref': 'BONUS-02-calculadora-plantilla-optima.xlsx:Calculadora:'
                   'B32:B35',
            'pruebas': pruebas, 'ok': all(x['ok'] for x in pruebas),
            'nota': 'el refuerzo es personal de ESE día y por eso no se suma '
                    'al FTE: es lo que la landing vende como «picos de '
                    'demanda» (COM-13)'}


def demos(carpeta, origen):
    fuera = {}
    for nombre, fn in (('grupo_c_06_ficha_y_na', _d06_ficha),
                       ('grupo_c_06_tendencia_dos_ultimos', _d06_tendencia),
                       ('grupo_c_07_vencimientos_30x4', _d07_vencimientos),
                       ('grupo_c_07_aviso_menor_de_edad', _d07_menor_edad),
                       ('grupo_c_bonus02_fte_y_ratio', _d_bonus_fte),
                       ('grupo_c_bonus02_dia_pico', _d_bonus_pico)):
        try:
            r = fn(carpeta)
        except Exception as e:                              # noqa: BLE001
            r = {'ok': False, 'error': '{}: {}'.format(type(e).__name__, e)}
        if r is not None:
            fuera[nombre] = r
    return fuera
