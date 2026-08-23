#!/usr/bin/env python3
"""
main.py — Orquestador del post-proceso v2.0 del Kit de Gestión de Personal
y Turnos.

    python3 main.py --dry-run [--solo motor] [--json informe.json]

`--solo` construye el kit ENTERO por defecto (`motor,a,b,c`). Se puede pedir una
pasada PARCIAL (`--solo motor`, `--solo a,b`) y entonces el informe la marca
como tal: los gates que miden trabajo de los grupos ausentes se declaran
`pendiente_de_grupo` en vez de `fallo`, porque suspender al motor por no haber
hecho el trabajo del grupo no dice nada. Lo que NO se relaja nunca es lo que sí
depende del motor: idempotencia, censo, protección y cache de sus propias
fórmulas.

Al revés también: pedir el kit ENTERO (el default) y que falte un `grupo_*.py`
es un FALLO con exit != 0. En el kit hermano el default era `motor` y el mismo
comando que documentaba la cabecera producía una copia con cero fórmulas nuevas
y cerraba con «TODO VERDE» (RD-29/RC-11): quien corriera el comando, leyera la
última línea y diera el visto bueno estaría aprobando un kit sin construir.

REGLA DURA: `astro-site/public/dl/kit-gestion-personal/` NO se toca. `--dry-run`
regenera una copia en el scratchpad y trabaja allí. Sin `--dry-run` el script
ABORTA salvo que se le pase `KIT_GESTION_PERSONAL_APPLY=1` en el entorno.

LOS MÓDULOS DE GRUPO PUEDEN NO EXISTIR. `--solo motor` corre sólo el §1;
`--solo a,b` carga los que encuentre y avisa —sin reventar— de los que no. Así
este orquestador es útil desde el primer commit y no hay que tocarlo cuando
`grupo_a.py`, `grupo_b.py` y `grupo_c.py` aterricen.

Qué hace, en este orden:
  1. Copia de trabajo (dry-run) desde `dl/kit-gestion-personal`.
  2. Por fichero: grupo.pre → motor.aplicar (§1) → grupo.post → motor.cerrar.
     Las columnas se insertan en `pre()` (antes de que el motor fije rangos);
     las filas se AÑADEN en `post()`.
  3. Idempotencia: repite el pipeline sobre un CLON del resultado y compara
     celda a celda (valor, formato, relleno, bloqueo, merges, DV, CF,
     protección, print_area, freeze). Debe dar 0 diferencias.
  4. `inject_cache.py` sobre los ficheros tocados. SIEMPRE al final: cualquier
     guardado posterior de openpyxl borraría el valor cacheado y el cliente
     vería las celdas de resultado en blanco en el visor del móvil.
  5. Verificación `data_only` de las fórmulas nuevas del registro + gate §1.5:
     ningún valor cacheado puede empezar por `#`.
  6. `censo-entregables.py --only <carpeta> --fail --quiet`.
  7. Gates de texto: pestañas citadas que no existen y citas legales obsoletas.
  8. Demostraciones con pycel (SPEC §5). pycel NO implementa `COUNTA` ni
     `MODE`; se usan `COUNTIF(rango,"<>")` e `INDEX`/`MATCH`. Y `MOD` SÍ evalúa
     con horas, con el ruido de coma flotante que obliga al `ROUND(...,2)`.

Térmica: todo en SERIE, un python a la vez. No hay builds ni navegador.
"""
import argparse
import contextlib
import datetime
import importlib
import json
import logging
import os
import shutil
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl                                            # noqa: E402

import motor                                               # noqa: E402

logging.disable(logging.CRITICAL)

AQUI = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(AQUI)
ROOT = os.path.dirname(os.path.dirname(SCRIPTS))
ORIGEN = os.path.join(ROOT, 'astro-site', 'public', 'dl',
                      'kit-gestion-personal')
SCRATCH = os.environ.get(
    'CLAUDE_SCRATCHPAD',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    '7340312f-b4fe-4aa1-b254-4b0c17c8375f/scratchpad')
DESTINO = os.path.join(SCRATCH, 'dryrun-v2', 'kit-gestion-personal')
IDEM = os.path.join(SCRATCH, 'dryrun-v2', 'kit-gestion-personal-idem')
INJECT = os.path.join(SCRIPTS, 'inject_cache.py')
CENSO = os.path.join(SCRIPTS, 'censo-entregables.py')

SPEC = 'scripts/productos-digitales/kit-gestion-personal-v2-SPEC.md §1 (motor)'
ENV_APPLY = 'KIT_GESTION_PERSONAL_APPLY'

#: Número de serie de Excel de una hora del día (fracción). pycel restando una
#: CADENA a una hora devuelve `#VALUE!`: las horas se inyectan como FRACCIÓN de
#: día, que es como las guarda el propio .xlsx.
def hora(hh, mm=0):
    return (hh * 60.0 + mm) / 1440.0


def log(msg):
    print(msg, flush=True)


# ==========================================================================
# Copia de trabajo
# ==========================================================================
def preparar_copia():
    if not os.path.isdir(ORIGEN):
        raise SystemExit('No existe el origen: ' + ORIGEN)
    if os.path.isdir(DESTINO):
        shutil.rmtree(DESTINO)
    os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
    shutil.copytree(ORIGEN, DESTINO)
    log('  copia de trabajo regenerada: ' + DESTINO)


# ==========================================================================
# Huella comparable e idempotencia
# ==========================================================================
def digest(path):
    """Huella de un .xlsx: valores, formatos, relleno, bloqueo, merges, DV, CF,
    protección, área de impresión y paneles. Es lo que compara la 2.ª pasada."""
    wb = openpyxl.load_workbook(path)
    fuera = {}
    for ws in wb.worksheets:
        celdas = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                relleno = None
                if c.fill is not None and c.fill.fill_type == 'solid':
                    relleno = str(c.fill.fgColor.rgb)
                celdas[c.coordinate] = (repr(c.value), c.number_format,
                                        relleno, bool(c.protection.locked))
        fuera[ws.title] = {
            'celdas': celdas,
            'merges': sorted(str(m) for m in ws.merged_cells.ranges),
            'dv': sorted('{}:{}:{}'.format(dv.type, dv.formula1, dv.sqref)
                         for dv in ws.data_validations.dataValidation),
            'cf': sorted('{}:{}'.format(cf.sqref, len(cf.rules))
                         for cf in ws.conditional_formatting),
            'prot': bool(ws.protection.sheet),
            'area': str(ws.print_area),
            'panel': '{}|{}|{}'.format(ws.freeze_panes, ws.print_title_rows,
                                       ws.print_title_cols),
        }
    fuera['__props__'] = {
        'celdas': {'title': (repr(wb.properties.title), '', None, False),
                   'subject': (repr(wb.properties.subject), '', None, False),
                   'keywords': (repr(wb.properties.keywords), '', None, False),
                   'creator': (repr(wb.properties.creator), '', None, False)},
        'merges': [], 'dv': [], 'cf': [], 'prot': False, 'area': 'None',
        'panel': ''}
    return fuera


def diff_digest(a, b, fichero):
    fuera = []
    for hoja in sorted(set(a) | set(b)):
        if hoja not in a or hoja not in b:
            fuera.append('{}:{}: hoja sólo en una pasada'.format(fichero, hoja))
            continue
        ha, hb = a[hoja], b[hoja]
        for k in ('merges', 'dv', 'cf', 'prot', 'area', 'panel'):
            if ha.get(k) != hb.get(k):
                fuera.append('{}:{}: cambia {} ({} → {})'
                             .format(fichero, hoja, k, ha.get(k), hb.get(k)))
        ca, cb = ha['celdas'], hb['celdas']
        for coord in sorted(set(ca) | set(cb)):
            if ca.get(coord) != cb.get(coord):
                fuera.append('{}:{}!{}: {} → {}'
                             .format(fichero, hoja, coord, ca.get(coord),
                                     cb.get(coord)))
    return fuera


# ==========================================================================
# Pipeline por fichero
# ==========================================================================
def procesar(carpeta, fname, grupos, informe_global):
    path = os.path.join(carpeta, fname)
    wb = openpyxl.load_workbook(path)
    cambios, registro_grupo = [], []
    motor.REGISTRO = []

    # Un grupo puede declarar `PROPIOS`: ficheros donde el motor §1 no aplica y
    # el grupo hace el trabajo entero (llamando él mismo a `motor.cerrar`).
    propio = any(fname in getattr(g, 'PROPIOS', []) for g in grupos)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []) and hasattr(g, 'pre'):
            g.pre(wb, fname, cambios)

    if not propio:
        motor.aplicar(wb, fname, cambios)

    for g in grupos:
        if fname in getattr(g, 'FICHEROS', []) and hasattr(g, 'post'):
            g.post(wb, fname, cambios, registro_grupo)

    gates = {}
    if not propio:
        gates = motor.cerrar(wb, fname, cambios)
    wb.save(path)

    registro = list(motor.REGISTRO)
    registro += [(h, c, f) for h, c, f in registro_grupo
                 if isinstance(f, str) and f.startswith('=')]
    informe_global.append({'fichero': fname, 'cambios': cambios,
                           'formulas_nuevas': len(registro), 'gates': gates})
    return registro, gates


def ficheros_a_tocar(grupos):
    nombres = list(motor.FICHEROS)
    for g in grupos:
        for f in getattr(g, 'FICHEROS', []):
            if f not in nombres:
                nombres.append(f)
    return nombres


# ==========================================================================
# Gates
# ==========================================================================
def inject_cache(carpeta, nombres):
    """SIEMPRE al final del pipeline: `inject_cache.py` escribe el valor
    cacheado directamente en el XML del zip, y cualquier `wb.save()` posterior
    de openpyxl lo borraría sin avisar."""
    fuera = []
    for n in nombres:
        r = subprocess.run([sys.executable, INJECT, os.path.join(carpeta, n)],
                           capture_output=True, text=True)
        fuera.append((n, r.stdout.strip(), r.returncode))
        log('    ' + (r.stdout.strip() or r.stderr.strip()[-200:]))
        if r.returncode != 0:
            log('    ERROR inject_cache: ' + r.stderr[-400:])
    return fuera


def verificar_cache(carpeta, registros):
    """Cada fórmula NUEVA debe tener valor cacheado.

    La celda sin valor cacheado NO se exime por su texto (que en este kit sería
    casi siempre, porque la guarda `=IF(...="","",…)` es el patrón de todo el
    §1.5): se le pregunta a pycel cuánto vale de verdad. Si pycel dice que vale
    la cadena vacía, el blanco es correcto; si devuelve cualquier otra cosa, es
    una fórmula que SÍ tenía valor y se quedó sin cachear —el cliente la vería
    en blanco en el visor del móvil— y va a `fallos`.
    """
    fallos, vacias, sinpycel, ok, no_evaluables = [], 0, 0, 0, 0
    for fname, registro in registros.items():
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path) or not registro:
            continue
        wbv = openpyxl.load_workbook(path, data_only=True)
        pendientes = []
        for hoja, coord, formula in registro:
            if hoja not in wbv.sheetnames:
                fallos.append('{}:{}!{}: hoja ausente'.format(fname, hoja,
                                                              coord))
                continue
            v = wbv[hoja][coord].value
            if v is not None:
                ok += 1
            elif any(f in formula.upper() for f in ('COUNTA(', 'MODE(',
                                                    'PMT(', 'IRR(')):
                sinpycel += 1
            else:
                pendientes.append((hoja, coord, formula))
        if not pendientes:
            continue
        try:
            xl = _pycel(path)
        except Exception as e:                                   # noqa: BLE001
            fallos.append('{}: no se pudo compilar con pycel para verificar '
                          'las {} fórmulas sin cache ({})'
                          .format(fname, len(pendientes), type(e).__name__))
            continue
        for hoja, coord, formula in pendientes:
            valor = _ev(xl, "'{}'!{}".format(hoja, coord))
            if isinstance(valor, str) and valor.startswith('ERR:'):
                no_evaluables += 1
            elif valor in ('', None):
                vacias += 1
            else:
                fallos.append('{}:{}!{}: SIN CACHE pero pycel la evalúa a {!r} '
                              '({})'.format(fname, hoja, coord, valor,
                                            formula[:60]))
    return {'con_valor': ok, 'vacias_verificadas_con_pycel': vacias,
            'sin_pycel': sinpycel, 'no_evaluables_por_pycel': no_evaluables,
            'fallos': fallos}


def gate_errores_cacheados(carpeta, nombres):
    """§1.5 — ninguna hoja recién descargada puede enseñar un `#¡DIV/0!`.

    Es el gate que mide DOM-02/COM-04: `06!Ficha Evaluación!C22` es hoy
    `=AVERAGE(C12:C21)` sobre diez celdas vacías, así que el documento que se
    firma delante del empleado abre con un error de Excel en la casilla de la
    nota media. Lo arregla `grupo_c` (§4) con `IF(COUNT(...)=0,"",…)`; el motor
    lo MIDE para que nadie dé por bueno un kit que lo siga teniendo.
    """
    fuera = {}
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        malos = motor.cacheados_con_error(path)
        if malos:
            fuera[n] = malos
    return fuera


def censo(carpeta):
    r = subprocess.run([sys.executable, CENSO, '--only', carpeta, '--fail',
                        '--quiet'], capture_output=True, text=True)
    log(r.stdout.strip())
    if r.returncode != 0:
        log(r.stderr.strip()[-1500:])
    salida = r.stdout.strip().splitlines()
    return {'exit': r.returncode, 'salida': salida[-6:],
            'stderr': r.stderr.strip().splitlines()[-12:]}


def gate_textos(carpeta, nombres):
    """Pestañas citadas que no existen + citas legales obsoletas + leyenda
    incoherente. Los tres se leen del fichero final, no del informe."""
    pest, citas, leyenda = {}, {}, {}
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        malas = motor.pestanas_citadas(wb)
        if malas:
            pest[n] = ['Instrucciones!{}: «{}» no existe (hojas: {})'
                       .format(f, txt, ', '.join(wb.sheetnames))
                       for f, txt in malas]
        obs = motor.citas_obsoletas(wb)
        if obs:
            citas[n] = obs
        inc = motor.leyenda_coherente(wb, n)
        if inc:
            leyenda[n] = inc
    return pest, citas, leyenda


# ==========================================================================
# Demostraciones con pycel (SPEC §5)
# ==========================================================================
def _pycel(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                                   # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)


def _set(xl, ref, valor):
    """pycel exige que la celda esté YA en el `cell_map` antes de escribirla
    (`AssertionError: Address … not found in the cell map`): se evalúa primero
    y luego se asigna."""
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            xl.evaluate(ref)
            xl.set_value(ref, valor)
            return True
        except Exception:                                        # noqa: BLE001
            return False


def _col_de(ws, fila, texto):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila, column=c).value
        if isinstance(v, str) and texto.lower() in v.lower():
            return openpyxl.utils.get_column_letter(c)
    return None


def demo_cruce_medianoche(carpeta):
    """SPEC §5 · 23:00→07:00 = 8,00 h y 19:00→01:30 = 6,50 h.

    La v1.1 hace `(D5-C5)*24` en `02-control-horas-extras.xlsx:Registro
    Horas:E5`: con un turno de noche da **−16 h**. Es el hallazgo DOM-03/COM-07
    y el que obliga al `MOD` con `ROUND` de §7-bis.2.
    """
    path = os.path.join(carpeta, '02-control-horas-extras.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Registro Horas']
    col_h = _col_de(ws, 4, 'Horas trabajadas')
    col_e = _col_de(ws, 4, 'entrada')
    col_s = _col_de(ws, 4, 'salida')
    col_p = _col_de(ws, 4, 'Pausa')
    if not (col_h and col_e and col_s):
        return {'ok': False, 'nota': 'no se localizan las columnas por su '
                                     'cabecera en la fila 4'}
    formula = ws['{}5'.format(col_h)].value
    xl = _pycel(path)
    pruebas = []
    for etiqueta, ent, sal, pausa, esperado in (
            ('turno de noche 23:00→07:00', hora(23), hora(7), None, 8.0),
            ('partido 19:00→01:30 con 0 h de pausa', hora(19), hora(1, 30),
             None, 6.5),
            ('partido 10:00→23:00 con 4 h de pausa', hora(10), hora(23), 4,
             9.0)):
        if pausa is not None and not col_p:
            pruebas.append({'caso': etiqueta, 'esperado': esperado,
                            'obtenido': None,
                            'nota': 'la columna «Pausa (h)» todavía no existe '
                                    '— la crea grupo_a (DOM-21)'})
            continue
        _set(xl, "'Registro Horas'!{}5".format(col_e), ent)
        _set(xl, "'Registro Horas'!{}5".format(col_s), sal)
        if col_p:
            _set(xl, "'Registro Horas'!{}5".format(col_p),
                 0 if pausa is None else pausa)
        v = _ev(xl, "'Registro Horas'!{}5".format(col_h))
        pruebas.append({'caso': etiqueta, 'esperado': esperado, 'obtenido': v,
                        'ok': isinstance(v, (int, float))
                        and abs(v - esperado) < 0.005})
    ok = all(p.get('ok') for p in pruebas)
    return {'ref': '02-control-horas-extras.xlsx:Registro Horas:{}5'
                   .format(col_h),
            'formula': formula, 'pruebas': pruebas, 'ok': ok,
            'pendiente_de_grupo': not ok,
            'nota': ('mientras la fórmula sea `(D5-C5)*24` el turno de noche '
                     'ficha −16 h y el partido no cabe en un solo par '
                     'entrada/salida (DOM-03/DOM-21)')}


def demo_guarda_contratadas(carpeta):
    """SPEC §5 · «H. contratadas» vacía → la hora extra se queda EN BLANCO.

    Hoy `02!Registro Horas!G5` es `=IF(E5<>"",MAX(0,E5-F5),"")`: con la columna
    F vacía, `MAX(0, 8-0)` declara extra la jornada ENTERA. A 12 €/h con recargo
    1,25 son 120 € por día y empleado que el cliente no debe (DOM-11/TEC-09).
    """
    path = os.path.join(carpeta, '02-control-horas-extras.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Registro Horas']
    col_x = _col_de(ws, 4, 'Horas extra')
    col_e = _col_de(ws, 4, 'entrada')
    col_s = _col_de(ws, 4, 'salida')
    col_c = _col_de(ws, 4, 'contratadas')
    if not (col_x and col_e and col_s and col_c):
        return {'ok': False, 'nota': 'no se localizan las columnas'}
    xl = _pycel(path)
    _set(xl, "'Registro Horas'!{}5".format(col_e), hora(9))
    _set(xl, "'Registro Horas'!{}5".format(col_s), hora(17))
    _set(xl, "'Registro Horas'!{}5".format(col_c), None)
    vacia = _ev(xl, "'Registro Horas'!{}5".format(col_x))
    _set(xl, "'Registro Horas'!{}5".format(col_c), 8)
    con8 = _ev(xl, "'Registro Horas'!{}5".format(col_x))
    ok = vacia in ('', None) and isinstance(con8, (int, float)) \
        and abs(con8) < 0.005
    return {'ref': '02-control-horas-extras.xlsx:Registro Horas:{}5'
                   .format(col_x),
            'formula': ws['{}5'.format(col_x)].value,
            'contratadas_vacia': vacia, 'contratadas_8h': con8,
            'ok': ok, 'pendiente_de_grupo': not ok,
            'nota': 'con la guarda puesta, «contratadas vacía» debe dar "" y '
                    'una jornada de 8 h contra 8 contratadas debe dar 0'}


def demo_ficha_vacia(carpeta):
    """SPEC §5 · ficha recién descargada → media EN BLANCO, no `#¡DIV/0!`."""
    path = os.path.join(carpeta, '06-evaluacion-desempeno.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Ficha Evaluación']
    xl = _pycel(path)
    media = _ev(xl, "'Ficha Evaluación'!C22")
    nivel = _ev(xl, "'Ficha Evaluación'!C23")
    ok = media in ('', None) and nivel in ('', None)
    return {'ref': '06-evaluacion-desempeno.xlsx:Ficha Evaluación:C22',
            'formula_media': ws['C22'].value, 'formula_nivel': ws['C23'].value,
            'media_con_ficha_vacia': str(media), 'nivel': str(nivel),
            'ok': ok, 'pendiente_de_grupo': not ok,
            'nota': 'es un documento que se FIRMA delante del empleado: no '
                    'puede abrir con un error de Excel en la nota media '
                    '(DOM-02/COM-04)'}


def demo_fte(carpeta):
    """SPEC §5 · el 03 y el BONUS-02 tienen que devolver los MISMOS 7 FTE con
    los datos por defecto (80 cubiertos/día, 2 servicios, casual).

    Hoy no: `03!Previsión por Servicio!B13` da 15 personas por turno y
    `BONUS-02!Calculadora!B23` da 19 en total — dos modelos de dimensionamiento
    distintos en el mismo kit, y el segundo con un ratio de coste laboral del
    99,7 % frente al 28-33 % de su propia hoja de referencia (DOM-09/DOM-10/
    COM-09).
    """
    fuera = {'esperado_fte': 7}
    p3 = os.path.join(carpeta, '03-coste-laboral-mensual.xlsx')
    if os.path.isfile(p3):
        wb = openpyxl.load_workbook(p3)
        ws = wb['Previsión por Servicio']
        fila = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and ('TOTAL personal' in v or 'FTE' in v):
                fila = r
                break
        if fila:
            xl = _pycel(p3)
            fuera['03_previsión'] = {
                'ref': '03-coste-laboral-mensual.xlsx:Previsión por '
                       'Servicio:B{}'.format(fila),
                'rotulo': ws.cell(row=fila, column=1).value,
                'valor': _ev(xl, "'Previsión por Servicio'!B{}".format(fila))}
    p9 = os.path.join(carpeta, 'BONUS-02-calculadora-plantilla-optima.xlsx')
    if os.path.isfile(p9):
        wb = openpyxl.load_workbook(p9)
        ws = wb['Calculadora']
        fila = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and ('Personal total' in v or 'FTE' in v):
                fila = r
                break
        if fila:
            xl = _pycel(p9)
            fuera['BONUS-02_calculadora'] = {
                'ref': 'BONUS-02-calculadora-plantilla-optima.xlsx:'
                       'Calculadora:B{}'.format(fila),
                'rotulo': ws.cell(row=fila, column=1).value,
                'valor': _ev(xl, "'Calculadora'!B{}".format(fila))}
    v3 = (fuera.get('03_previsión') or {}).get('valor')
    v9 = (fuera.get('BONUS-02_calculadora') or {}).get('valor')
    fuera['coinciden'] = (v3 == v9)
    fuera['ok'] = (v3 == 7 and v9 == 7)
    fuera['pendiente_de_grupo'] = not fuera['ok']
    fuera['nota'] = ('un solo modelo de dimensionamiento: cubiertos por '
                     'SERVICIO → personal por servicio → presencias/día → '
                     'horas semanales → FTE (§3, grupo_b)')
    return fuera


def demo_descanso(carpeta):
    """SPEC §5 · T seguido de M dispara la alerta de descanso (23:00→07:00 son
    8 h, por debajo de las 12 del art. 34.3 ET) y M seguido de M no.

    Las cuatro alertas del cuadrante las construye `grupo_a` (§2): hasta
    entonces, `01!Cuadrante Semanal!J` sólo mira si se pasa de 40 h a la semana.
    """
    path = os.path.join(carpeta, '01-cuadrante-turnos-semanal.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Cuadrante Semanal']
    col = _col_de(ws, 5, 'Descanso entre')
    if not col:
        return {'ok': False, 'pendiente_de_grupo': True,
                'ref': '01-cuadrante-turnos-semanal.xlsx:Cuadrante Semanal:K5',
                'columnas_hoy': [ws.cell(row=5, column=c).value
                                 for c in range(1, ws.max_column + 1)],
                'nota': 'la columna «Descanso entre jornadas» todavía no '
                        'existe — la crea grupo_a (DOM-05/COM-01). Hoy la '
                        'única alerta del cuadrante es «>40 h» y la landing '
                        'promete cuatro.'}
    xl = _pycel(path)
    pruebas = []
    for etiqueta, lunes, martes, dispara in (
            ('T seguido de M (8 h de descanso)', 'T', 'M', True),
            ('M seguido de M (24 h de descanso)', 'M', 'M', False)):
        _set(xl, "'Cuadrante Semanal'!B6", lunes)
        _set(xl, "'Cuadrante Semanal'!C6", martes)
        v = _ev(xl, "'Cuadrante Semanal'!{}6".format(col))
        salta = bool(v) and v not in ('', None)
        pruebas.append({'caso': etiqueta, 'valor': v,
                        'esperaba_alerta': dispara, 'ok': salta == dispara})
    return {'ref': '01-cuadrante-turnos-semanal.xlsx:Cuadrante Semanal:{}6'
                   .format(col),
            'pruebas': pruebas, 'ok': all(p['ok'] for p in pruebas),
            'pendiente_de_grupo': not all(p['ok'] for p in pruebas)}


def demo_progreso_onboarding(carpeta):
    """SPEC §4/§5 · el checklist recién descargado tiene que marcar 0 %.

    Hoy `04!Checklist Onboarding!C68` es `=COUNTIF(F7:F65,"✓")` y ese rango se
    traga las CABECERAS de las secciones 2ª a 5ª —`F19`, `F30`, `F42` y `F57`
    valen literalmente «✓»—, así que un onboarding sin empezar abre en «4 de
    51 completadas» y un 8,51 % (DOM-01/TEC-02/TEC-13, y es la mejora estrella
    del changelog v1.1).
    """
    path = os.path.join(carpeta, '04-onboarding-nuevo-empleado.xlsx')
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb['Checklist Onboarding']
    fila_c = fila_t = fila_p = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if not isinstance(v, str):
            continue
        if 'completadas' in v.lower():
            fila_c = r
        elif v.lower().startswith('total tareas'):
            fila_t = r
        elif 'progreso' in v.lower():
            fila_p = r
    xl = _pycel(path)
    completadas = _ev(xl, "'Checklist Onboarding'!C{}".format(fila_c)) \
        if fila_c else None
    total = _ev(xl, "'Checklist Onboarding'!C{}".format(fila_t)) \
        if fila_t else None
    progreso = _ev(xl, "'Checklist Onboarding'!C{}".format(fila_p)) \
        if fila_p else None
    tramos = [{'cabecera': hdr, 'primera': r0, 'ultima': r1,
               'tareas': r1 - r0 + 1}
              for _t, hdr, r0, r1 in motor.secciones_04(ws)]
    ok = (completadas == 0)
    return {'ref': '04-onboarding-nuevo-empleado.xlsx:Checklist '
                   'Onboarding:C{}'.format(fila_c or '?'),
            'formula_completadas': ws['C{}'.format(fila_c)].value
            if fila_c else None,
            'completadas_recien_descargado': completadas,
            'total_declarado': total, 'progreso': progreso,
            'tramos_detectados': tramos,
            'tareas_detectadas': sum(t['tareas'] for t in tramos),
            'ok': ok, 'pendiente_de_grupo': not ok,
            'nota': 'los contadores tienen que acotarse por TRAMO '
                    '(COUNTIF por sección) y las cabeceras pasar de «✓» a '
                    '«Hecho» — se hacen las dos cosas (§4)'}


def demo_leyenda(carpeta):
    """§1.1 — la DV de códigos ofrece los 8 de jornada y los 4 de ausencia, y
    el permiso es `PE`, nunca `P`."""
    fuera = {}
    for fname, hoja, esperado in (
            ('01-cuadrante-turnos-semanal.xlsx', 'Cuadrante Semanal',
             motor.DV_JORNADA),
            ('05-planificacion-vacaciones.xlsx', 'Calendario Anual',
             motor.DV_AUSENCIA)):
        path = os.path.join(carpeta, fname)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        listas = [str(dv.formula1) for dv in ws.data_validations.dataValidation
                  if isinstance(getattr(dv, 'formula1', None), str)
                  and dv.formula1.startswith('"')]
        leyenda = [ws.cell(row=3, column=c).value
                   for c in range(2, ws.max_column + 1)
                   if isinstance(ws.cell(row=3, column=c).value, str)]
        fuera[fname] = {
            'ref': '{}:{}:B3'.format(fname, hoja),
            'dv_encontradas': listas,
            'dv_esperada': '"{}"'.format(esperado),
            'leyenda_fila_3': leyenda,
            'ok': ('"{}"'.format(esperado) in listas
                   and len(listas) == 1)}
    fuera['ok'] = all(v.get('ok') for k, v in fuera.items() if k != 'ok')
    return fuera


def demo_proteccion(carpeta, nombres):
    """§1.6 — hoja a hoja: protegida sin contraseña y con las verdes abiertas."""
    fuera = []
    for n in nombres:
        path = os.path.join(carpeta, n)
        if not os.path.isfile(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            verdes = sum(1 for row in ws.iter_rows() for c in row
                         if motor.es_verde(c))
            abiertas = sum(1 for row in ws.iter_rows() for c in row
                           if motor.es_verde(c) and not c.protection.locked)
            fuera.append({'ref': '{}:{}'.format(n, ws.title),
                          'protegida': bool(ws.protection.sheet),
                          'con_password': bool(ws.protection.password),
                          'verdes': verdes, 'verdes_desbloqueadas': abiertas})
    fallos = ['{}: {} verdes y sólo {} desbloqueadas'
              .format(d['ref'], d['verdes'], d['verdes_desbloqueadas'])
              for d in fuera if d['verdes'] != d['verdes_desbloqueadas']]
    fallos += ['{}: protegida CON contraseña'.format(d['ref'])
               for d in fuera if d['con_password']]
    return {'hojas': fuera, 'fallos': fallos}


# ==========================================================================
# main
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description='Post-proceso v2.0 del Kit de Gestión de Personal y Turnos')
    ap.add_argument('--dry-run', action='store_true',
                    help='trabaja sobre una copia en el scratchpad')
    ap.add_argument('--solo', default='motor,a,b,c',
                    help='qué aplicar: motor,a,b,c (por defecto) | motor | '
                         'a,b … (el motor va siempre)')
    ap.add_argument('--json', default=None, help='ruta del informe JSON')
    ap.add_argument('--sin-idempotencia', action='store_true')
    ap.add_argument('--sin-demos', action='store_true')
    args = ap.parse_args()

    if not args.dry_run and os.environ.get(ENV_APPLY) != '1':
        raise SystemExit(
            'ABORTADO: sin --dry-run este script escribiría en ' + ORIGEN +
            ', que la SPEC declara intocable hasta que la ronda 2 dé verde. '
            'Usa --dry-run (o ' + ENV_APPLY + '=1 si eres el orquestador y ya '
            'tienes el visto bueno).')

    carpeta = DESTINO if args.dry_run else ORIGEN
    respaldo = None
    if args.dry_run:
        preparar_copia()
    else:
        # En producción se escribe IN PLACE sobre los entregables. Si algo
        # revienta a mitad del bucle, los xlsx quedan a medio transformar. El
        # respaldo va al SCRATCHPAD, nunca junto a los entregables: una carpeta
        # `.bak` dentro de `public/dl/` se publicaría.
        respaldo = os.path.join(
            SCRATCH, 'kit-gestion-personal.bak-'
            + datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
        os.makedirs(os.path.dirname(respaldo), exist_ok=True)
        shutil.copytree(ORIGEN, respaldo)
        log('  respaldo previo de los entregables: ' + respaldo)

    grupos, pedidos, ausentes = [], [], []
    for letra in [g.strip().lower() for g in args.solo.split(',') if g.strip()]:
        if letra == 'motor':
            continue
        # `--solo motor,grupo_a` y `--solo motor,a` son la misma cosa: sin esto,
        # la primera forma buscaría un módulo `grupo_grupo_a` y el grupo se
        # declararía «ausente» sin que nada fallara — un verde falso, que es el
        # peor resultado posible de un gate.
        if letra.startswith('grupo_'):
            letra = letra[6:]
        pedidos.append(letra)
        try:
            grupos.append(importlib.import_module('grupo_' + letra))
            log('  grupo_{} cargado'.format(letra))
        except ImportError:
            ausentes.append('grupo_' + letra)
            log('  grupo_{} no existe todavía — se omite'.format(letra))

    #: `completo` = se pidieron los tres grupos Y los tres están. Con `--solo
    #: motor` NO se pidió ninguno: la pasada es parcial DECLARADA y los gates
    #: que miden trabajo de grupo informan en vez de suspender.
    completo = set(pedidos) >= {'a', 'b', 'c'} and not ausentes
    parcial_declarada = not set(pedidos) & {'a', 'b', 'c'}

    nombres = ficheros_a_tocar(grupos)
    informe_ficheros, registros, gates_fichero = [], {}, {}
    log('\n== 1/7 · post-proceso de {} ficheros =='.format(len(nombres)))
    for fname in nombres:
        registros[fname], gates_fichero[fname] = procesar(
            carpeta, fname, grupos, informe_ficheros)
        log('  {}: {} fórmulas nuevas'.format(fname, len(registros[fname])))

    # ---- idempotencia ---------------------------------------------------
    idem = {'ejecutada': False}
    if not args.sin_idempotencia:
        log('\n== 2/7 · idempotencia (2.ª pasada sobre un clon) ==')
        if os.path.isdir(IDEM):
            shutil.rmtree(IDEM)
        shutil.copytree(carpeta, IDEM)
        antes = dict((n, digest(os.path.join(carpeta, n))) for n in nombres)
        for fname in nombres:
            procesar(IDEM, fname, grupos, [])
        difs = []
        for n in nombres:
            difs += diff_digest(antes[n], digest(os.path.join(IDEM, n)), n)
        idem = {'ejecutada': True, 'diferencias': len(difs),
                'detalle': difs[:40]}
        log('  diferencias 1.ª vs 2.ª pasada: {}'.format(len(difs)))
        for d in difs[:10]:
            log('    ' + d)

    # ---- cache ----------------------------------------------------------
    log('\n== 3/7 · inject_cache (al final del todo) ==')
    cache = inject_cache(carpeta, nombres)

    log('\n== 4/7 · verificación data_only + gate «ningún cacheado con #» ==')
    ver = verificar_cache(carpeta, registros)
    log('  con valor: {} · "" VERIFICADA con pycel: {} · sin pycel: {} · '
        'no evaluables: {} · fallos: {}'
        .format(ver['con_valor'], ver['vacias_verificadas_con_pycel'],
                ver['sin_pycel'], ver['no_evaluables_por_pycel'],
                len(ver['fallos'])))
    for f in ver['fallos'][:10]:
        log('    ' + f)
    errores = gate_errores_cacheados(carpeta, nombres)
    n_err = sum(len(v) for v in errores.values())
    log('  valores cacheados que empiezan por «#»: {}'.format(n_err))
    for n, lineas in errores.items():
        for l in lineas[:4]:
            log('    {}: {}'.format(n, l))

    log('\n== 5/7 · censo-entregables --fail ==')
    cen = censo(carpeta)

    log('\n== 6/7 · gates de texto (pestañas, citas legales, leyenda) ==')
    pest, citas, leyenda = gate_textos(carpeta, nombres)
    log('  pestañas citadas inexistentes: {} ficheros'.format(len(pest)))
    for n, lineas in pest.items():
        for l in lineas:
            log('    {}: {}'.format(n, l))
    log('  citas legales obsoletas: {} en {} ficheros'
        .format(sum(len(v) for v in citas.values()), len(citas)))
    for n, lineas in list(citas.items())[:9]:
        for l in lineas[:3]:
            log('    {}: {}'.format(n, l))
    log('  DV con códigos fuera de la leyenda única: {}'
        .format(sum(len(v) for v in leyenda.values())))
    for n, lineas in leyenda.items():
        for l in lineas[:3]:
            log('    {}: {}'.format(n, l))

    log('\n== 7/7 · demostraciones (pycel) ==')
    demos = {}
    if not args.sin_demos:
        demos = {
            'cruce_medianoche_02': demo_cruce_medianoche(carpeta),
            'guarda_h_contratadas_02': demo_guarda_contratadas(carpeta),
            'ficha_vacia_06': demo_ficha_vacia(carpeta),
            'fte_03_vs_bonus02': demo_fte(carpeta),
            'descanso_entre_jornadas_01': demo_descanso(carpeta),
            'progreso_onboarding_04': demo_progreso_onboarding(carpeta),
            'leyenda_unica': demo_leyenda(carpeta),
            'proteccion': demo_proteccion(carpeta, nombres),
        }
        for g in grupos:
            if hasattr(g, 'demos'):
                propias = g.demos(carpeta, ORIGEN)
                demos.update(propias)
                log('  demostraciones de {}: {} bloques'
                    .format(g.__name__, len(propias)))
        for k, v in sorted(demos.items()):
            if isinstance(v, dict) and 'ok' in v:
                log('  {}: {}'.format(k, 'OK' if v['ok'] else
                                      ('PENDIENTE DE GRUPO'
                                       if v.get('pendiente_de_grupo')
                                       else 'FALLA')))

    # ---- veredicto ------------------------------------------------------
    #: Lo que el MOTOR controla y no se relaja nunca.
    fallos = []
    if idem.get('diferencias'):
        fallos.append('idempotencia: {} diferencias'
                      .format(idem['diferencias']))
    if ver['fallos']:
        fallos.append('cache: {} fórmulas sin valor'.format(len(ver['fallos'])))
    if cen['exit'] != 0:
        fallos.append('censo-entregables --fail devolvió {}'
                      .format(cen['exit']))
    if demos.get('proteccion'):
        fallos += demos['proteccion']['fallos']
    if leyenda:
        fallos.append('leyenda única: {} DV ofrecen códigos que la leyenda no '
                      'explica'.format(sum(len(v) for v in leyenda.values())))

    #: Lo que miden los gates del CONTENIDO. En una pasada completa bloquean;
    #: en una parcial declarada se listan como pendientes, que es lo honesto:
    #: el motor no escribe las fórmulas del §2/§3/§4 y suspenderle por no
    #: haberlas escrito no informa de nada. Lo que NO se hace nunca es
    #: esconderlos: van al informe y a la salida en los dos casos.
    pendientes_contenido = []
    if n_err:
        pendientes_contenido.append(
            '{} valores cacheados empiezan por «#» (§1.5): {}'
            .format(n_err, '; '.join('{}:{}'.format(n, v[0])
                                     for n, v in list(errores.items())[:3])))
    if pest:
        pendientes_contenido.append(
            'gate de pestañas: {} ficheros citan pestañas inexistentes'
            .format(len(pest)))
    fallidas = sorted(k for k, v in demos.items()
                      if isinstance(v, dict) and v.get('ok') is False)
    if fallidas:
        pendientes_contenido.append(
            'demostraciones sin pasar: ' + ', '.join(fallidas))
    if citas:
        pendientes_contenido.append(
            '{} citas legales obsoletas vivas en {} ficheros (11 h de '
            'descanso, ×1,75/×2,0 como ley, SS al 30 %, alérgenos propios)'
            .format(sum(len(v) for v in citas.values()), len(citas)))

    if completo:
        fallos += pendientes_contenido
    elif not parcial_declarada:
        fallos.append(
            'pasada PARCIAL no declarada: grupos pedidos {} · ausentes {}. El '
            'resultado NO es la v2.0 y no se puede dar por bueno.'
            .format(pedidos or ['ninguno'], ausentes or ['ninguno']))
        fallos += pendientes_contenido

    informe = {
        'producto': motor.PID,
        'version': motor.VERSION,
        'spec': SPEC,
        'rol': 'motor',
        'fecha': datetime.datetime.now().isoformat(timespec='seconds'),
        'modo': 'dry-run' if args.dry_run else 'produccion',
        'pasada': ('completa' if completo else
                   ('parcial declarada (--solo {})'.format(args.solo)
                    if parcial_declarada else 'parcial con grupos ausentes')),
        'carpeta_origen': ORIGEN,
        'carpeta_trabajo': carpeta,
        'grupos_pedidos': pedidos,
        'grupos_ausentes': ausentes,
        'ficheros': informe_ficheros,
        'gates': {
            'idempotencia': idem,
            'inject_cache': [{'fichero': n, 'salida': s, 'exit': rc}
                             for n, s, rc in cache],
            'data_only_formulas_nuevas': ver,
            'valores_cacheados_con_error': errores,
            'censo_entregables': cen,
            'pestanas_citadas_inexistentes': pest,
            'citas_legales_obsoletas': citas,
            'leyenda_incoherente': leyenda,
        },
        'demostraciones': demos,
        'pendiente_de_grupo': pendientes_contenido,
        'fallos': fallos,
        'exit': 1 if fallos else 0,
        'respaldo': respaldo,
    }
    if args.json:
        destino_json = os.path.abspath(args.json)
        if os.path.dirname(destino_json):
            os.makedirs(os.path.dirname(destino_json), exist_ok=True)
        with open(destino_json, 'w', encoding='utf-8') as fh:
            json.dump(informe, fh, ensure_ascii=False, indent=1, default=str)
        log('\ninforme → ' + destino_json)

    if pendientes_contenido and not completo:
        log('\nPENDIENTE DE GRUPO (no suspende una pasada parcial declarada):')
        for p in pendientes_contenido:
            log('  · ' + p)
    log('\n' + ('FALLOS:\n  ' + '\n  '.join(fallos) if fallos
                else 'MOTOR VERDE (idempotencia, cache, censo, protección, '
                     'leyenda única)'))
    if respaldo:
        log('  respaldo de los entregables previos en ' + respaldo
            + ('  (BÓRRALO sólo cuando compruebes el resultado)' if not fallos
               else '  ← RESTAURA DESDE AQUÍ: la pasada acabó con fallos'))
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
