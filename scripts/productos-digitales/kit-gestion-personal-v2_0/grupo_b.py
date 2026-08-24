#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
grupo_b.py — Kit de Gestión de Personal y Turnos v2.0

Reparto del orquestador (2026-08-24): este grupo construye TRES ficheros —

  · `03-coste-laboral-mensual.xlsx`   (SPEC §3)
  · `04-onboarding-nuevo-empleado.xlsx` (SPEC §4, bloque «04»)
  · `05-planificacion-vacaciones.xlsx`  (SPEC §4, bloque «05»)

`BONUS-02-calculadora-plantilla-optima.xlsx`, que la SPEC empareja con el 03
dentro del §3, NO está en este reparto: lo construye otro grupo. Aquí se deja
el 03 dando **7 FTE** con los datos por defecto y el modelo entero escrito en
celdas visibles (§3), que es lo que `main.demo_fte` compara contra el bonus.

Contrato con `main.py`: se exponen `FICHEROS`, `pre(wb, fname, cambios)` y
`demos(carpeta, origen)`. Todo el trabajo va en `pre()` —antes de
`motor.aplicar` y, por tanto, mucho antes de `motor.cerrar`— por dos razones
medidas:

  1. `motor.colores_seccion_04` corre dentro de `motor.aplicar`, es decir ENTRE
     `pre` y `post`, y detecta los cinco tramos del 04 con `secciones_04()`. Si
     las tres tareas nuevas (DOM-15) se insertaran en `post`, el motor habría
     pintado los tramos VIEJOS y las filas nuevas se quedarían sin color.
  2. `motor.cerrar` fija formatos, verde, DV, formato condicional, área de
     impresión y protección sobre el layout FINAL. Cuanto antes esté ese layout,
     menos casos especiales.

Idempotencia (2.ª pasada = 0 diferencias): las hojas que cambian de GEOMETRÍA
—el calendario del 05 pasa de 12 columnas a 55— se **reconstruyen enteras**
(`wb.remove` + `create_sheet` en el mismo índice) en vez de parchearse; las que
sólo cambian de contenido se sobrescriben celda a celda. Las DV propias se
borran antes de reescribirse (una `DataValidation` añadida dos veces son dos
entradas distintas en el `digest` de `main.py`), y lo mismo con el formato
condicional propio.
"""

import copy
import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import motor

FICHEROS = [
    '03-coste-laboral-mensual.xlsx',
    '04-onboarding-nuevo-empleado.xlsx',
    '05-planificacion-vacaciones.xlsx',
]

#: Marca de las DV que escribe ESTE grupo. Empieza por `motor.MARCA_DV` a
#: propósito: así `motor._limpiar_dv(ws)` —que reconoce las suyas por ese
#: prefijo— también se lleva las mías cuando limpia una hoja, y no quedan dos
#: validaciones sobre la misma celda (Excel aplica la primera que encuentra).
MARCA_B = motor.MARCA_DV + 'b'

# ==========================================================================
# Utilidades locales
# ==========================================================================


def _merges(ws, deseadas):
    """Fija EXACTAMENTE las combinaciones de la hoja. Idempotente por
    construcción: deshace todas y rehace la lista pedida."""
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for m in deseadas:
        ws.merge_cells(m)


def _anchos(ws, mapa):
    for letra, ancho in mapa.items():
        ws.column_dimensions[letra].width = ancho


def _vaciar(ws, min_row=1, max_row=None, fills=True):
    """Vacía valor, relleno y formato de las celdas que YA EXISTEN.

    Se recorre `iter_rows()` en vez de `ws.cell(r, c)`: pedir una celda que no
    existe la CREA y con ella agranda `ws.max_row`/`max_column`, y el
    `print_area` que fija `motor.cerrar` sale con filas y columnas en blanco
    dentro. Y se limpia el RELLENO, no sólo el valor: la v1.1 pinta de verde
    las columnas que en ella eran de entrada, y al reordenar las columnas ese
    verde se queda debajo de una columna ya CALCULADA diciéndole al cliente
    que puede sobrescribir la fórmula (`03!Nóminas!F` era «Horas Contratadas»
    y pasa a ser la cotización; `05!Solicitudes!E` era «Días Usados YTD» y
    pasa a leerse por INDEX/MATCH). Medido el 2026-08-24.
    """
    for fila in ws.iter_rows(min_row=min_row,
                             max_row=max_row or ws.max_row):
        for cel in fila:
            if isinstance(cel.value, str) or cel.value is not None:
                cel.value = None
            if fills:
                cel.fill = PatternFill()
                cel.number_format = 'General'


def _rotulo(ws, fila, texto, col=1, negrita=False):
    cel = ws.cell(row=fila, column=col, value=texto)
    if negrita:
        cel.font = Font(bold=True)
    return cel


def _formula(ws, coord, formula, fmt=None):
    """Escribe una fórmula y la REGISTRA para que `main.py` compruebe que
    quedó con valor cacheado (`verificar_cache`)."""
    cel = ws[coord]
    cel.value = formula
    if fmt:
        cel.number_format = fmt
    motor._reg(ws, coord, formula)
    return cel


def _verde(ws, coord, valor=None, fmt=None):
    cel = ws[coord]
    if valor is not None:
        cel.value = valor
    if fmt:
        cel.number_format = fmt
    motor.marcar_verde(ws, coord)
    return cel


def _dv(ws, rango, valores, titulo, prompt, error, marca=MARCA_B,
        tipo='list', formula=None):
    """DV marcada para poder retirarla en la 2.ª pasada. `tipo='custom'` +
    `formula` para las validaciones que NO pueden ser una lista inline (ver
    la fila de temporada del calendario del 05)."""
    dv = DataValidation(
        type=tipo,
        formula1=formula if tipo != 'list' else '"{}"'.format(
            ','.join(valores)),
        allow_blank=True, showErrorMessage=True, errorTitle=titulo,
        error=error, errorStyle='stop', showInputMessage=True,
        promptTitle='{} · {}'.format(marca, titulo), prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(rango)
    return dv


def _cf(ws, rango, vocabulario):
    """Semáforo propio, limpiando antes el rango: `motor.aplicar_cf` sólo
    limpia los rangos que gobierna ÉL, así que una regla del grupo escrita dos
    veces se duplicaría en cada pasada."""
    motor._limpiar_cf(ws, {rango})
    return motor.semaforo(ws, rango, vocabulario)


def _instrucciones(ws, lineas, zona=32):
    """Reescribe el cuerpo de `Instrucciones` (columna B en este kit).

    Se limpia una ZONA fija y se escribe desde la fila 2, dejando SIEMPRE
    contenido en la última fila escrita: el bloque de bio + versión que añade
    `motor.bio_y_version` se ancla en sí mismo, así que mientras mi zona no lo
    pise, la 2.ª pasada lo reescribe en su sitio y no se desplaza.
    """
    col = motor.col_instrucciones(ws)
    for r in range(2, zona + 1):
        ws.cell(row=r, column=col).value = None
    fila = 2
    for texto in lineas:
        if texto is None:
            fila += 1
            continue
        cel = ws.cell(row=fila, column=col, value=texto)
        cel.alignment = Alignment(wrap_text=True, vertical='top')
        if fila == 2:
            cel.font = Font(bold=True, size=13)
        fila += 1
    return fila - 1


def _cabecera(ws, fila, textos, col0=1):
    """Cabecera oscura, del mismo estilo que ya usa la v1.1."""
    for i, txt in enumerate(textos):
        cel = ws.cell(row=fila, column=col0 + i, value=txt)
        cel.fill = PatternFill('solid', fgColor=motor.CAB)
        cel.font = Font(bold=True, color='FFFFFF', size=10)
        cel.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)


def _seccion(ws, fila, texto, ancho, color='ECEFF1'):
    cel = ws.cell(row=fila, column=1, value=texto)
    cel.font = Font(bold=True, size=11)
    cel.fill = PatternFill('solid', fgColor=color)
    for c in range(2, ancho + 1):
        ws.cell(row=fila, column=c).fill = PatternFill('solid', fgColor=color)


def _pie(ws, fila, ancho):
    cel = ws.cell(row=fila, column=1, value=motor.PIE)
    cel.font = Font(size=8, color='888888')
    cel.alignment = Alignment(horizontal='center')
    return '{}{}:{}{}'.format('A', fila, get_column_letter(ancho), fila)


# ==========================================================================
# 03 · COSTE LABORAL MENSUAL  (SPEC §3)
# ==========================================================================
#: Los 6 tipos de negocio de la tabla de referencia del propio fichero
#: (`03-coste-laboral-mensual.xlsx:Ratio Coste Laboral:A14:A19`), con sus dos
#: columnas NUMÉRICAS nuevas (objetivo / máximo aceptable). No son cifras
#: inventadas: son el extremo superior de los rangos que la hoja ya imprime en
#: texto desde la v1.1 —«25-28 %» → 0,28 objetivo; «22-30 %» → 0,30 aceptable—,
#: así que el semáforo y la tabla dejan de contradecirse (COM-11).
TIPOS_NEGOCIO = [
    ('Fast Casual / Comida Rápida', '25-28%', '22-30%', 0.28, 0.30),
    ('Restaurante Casual',          '28-33%', '25-35%', 0.33, 0.35),
    ('Fine Dining / Alta Cocina',   '35-40%', '33-42%', 0.40, 0.42),
    ('Catering / Eventos',          '30-35%', '28-38%', 0.35, 0.38),
    ('Cafetería / Brunch',          '28-32%', '25-35%', 0.32, 0.35),
    ('Bar / Cocktails',             '25-30%', '22-33%', 0.30, 0.33),
]

FILA_TOT_03 = 4 + motor.CAPACIDAD + 1          # 35 · TOTALES de Nóminas

#: RD-20 · 52 semanas menos ~4,3 de vacaciones (30 días naturales, art. 38 ET)
#: y ~2 de festivos (14 días). Va en celda verde: cada convenio tiene lo suyo.
SEMANAS_EFECTIVAS = 46.5


def _nota_dv(ws, coord, titulo, prompt):
    """Sólo el mensaje de entrada (nota de parámetro), sin restringir valores:
    es lo que hace `motor.parametro` con los del catálogo §1.4 y lo que
    necesitan los parámetros propios de un fichero."""
    dv = DataValidation(type='decimal', operator='greaterThan', formula1='0',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle=titulo,
                        error='Escribe un número mayor que cero.',
                        errorStyle='stop', showInputMessage=True,
                        promptTitle='{} · {}'.format(MARCA_B, titulo),
                        prompt=prompt)
    ws.add_data_validation(dv)
    dv.add(coord)
    return dv


def _n03_nominas(wb, cambios):
    """§3 · DOM-08/DOM-31/TEC-14/TEC-26.

    Lo que había: `D5='=IF(C5<>"",C5*0.30,"")'` repetido en 20 filas, con el
    tipo de cotización ESCRITO DENTRO de la fórmula (21 veces contando la
    cabecera `D4='SS Empresa (30%)'`), sin pagas extra y con una columna
    «Horas Contratadas» que no alimentaba nada.
    """
    ws = wb['Nóminas']

    # 1) 20 → 30 empleados. La cola (TOTALES en 25, pie en 27) baja 10 filas y
    #    `=SUM(C5:C24)` se ESTIRA a `C34` (lo hace `motor._corre_cola`).
    delta = motor.expandir_filas(ws, 24, 4 + motor.CAPACIDAD, cola=(25, 27))
    if delta:
        cambios.append('03:Nóminas!5:34: bloque de empleados 20 → {} '
                       '(§1.3, DOM-32)'.format(motor.CAPACIDAD))

    _merges(ws, ['A1:I1', 'A3:I3',
                 'A{0}:I{0}'.format(FILA_TOT_03 + 2)])
    ws['A1'] = 'Nóminas — Mes: _______________'
    # el subtítulo de marca baja a la fila 3: la 2 pasa a ser el parámetro de
    # cotización (§3, «verde C2»). Se escribe LITERAL, no copiando lo que haya
    # en A2: en la 2.ª pasada A2 ya es el rótulo del parámetro y copiarlo
    # sustituiría el subtítulo por «Tipo de SS a cargo de la empresa (%):».
    ws['A3'] = 'AI Chef Pro · aichef.pro — Kit Gestión de Personal y Turnos'
    ws['A2'] = None
    ws['B2'] = None
    ws['C2'] = None

    # 2) el 0,30 sale de las 21 fórmulas y pasa a UNA celda verde (§1.4)
    motor._limpiar_dv(ws)
    coord = motor.parametro(ws, 2, 'ss_empresa', col_rotulo=1, col_valor=3)
    _formula(ws, 'D2', '="El coste total de cada empleado incluye un "'
                       '&TEXT($C$2,"0%")&" de cotización empresarial sobre el '
                       'bruto ya prorrateado"')

    # RD-20 · el «Coste/hora» dividía entre las horas contratadas de las 52
    # semanas del año, contando como trabajadas los 30 días naturales de
    # vacaciones (05!'Saldo Vacaciones'!B2, el propio kit) y los ~14 festivos:
    # se pagan y no se producen. El resultado salía entre un 10 y un 13 %
    # corto, y ese error se propagaba a la tarifa del 02 y de ahí al coste de
    # las horas extra. Las semanas EFECTIVAS pasan a celda verde, como manda
    # §1.4: dependen del convenio de cada uno.
    _rotulo(ws, 2, 'Semanas efectivas al año:', col=5, negrita=True)
    _verde(ws, 'F2', SEMANAS_EFECTIVAS, motor.FMT_DEC1)
    _nota_dv(ws, 'F2', 'Semanas efectivas al año',
             'Las 52 del año MENOS las de vacaciones y festivos, que se pagan '
             'y no se trabajan: con 30 días naturales de vacaciones (art. 38 '
             'ET) y unos 14 festivos salen ~46,5. Es el divisor del '
             '«Coste/hora»: con 52 el coste por hora REALMENTE trabajada sale '
             'un 10-13 % corto, y ese número es el que se lleva a la tarifa '
             'del fichero 02.')
    cambios.append('03:Nóminas!{}: tipo de cotización empresarial en celda '
                   'verde (33 %) — antes iba dentro de 20 fórmulas y de la '
                   'cabecera D4 (DOM-08/TEC-14)'.format(coord))

    # 3) cabecera de 6 → 9 columnas
    _cabecera(ws, 4, ['Empleado', 'Puesto', 'Salario Bruto Mes (€)',
                      'Nº de pagas/año', 'Bruto prorrateado/mes (€)',
                      'Cotización SS empresa (€)', 'Coste total/mes (€)',
                      'Horas contratadas/semana', 'Coste/hora (€)'])
    _anchos(ws, {'A': 24, 'B': 18, 'C': 18, 'D': 14, 'E': 20, 'F': 20,
                 'G': 18, 'H': 20, 'I': 14})

    # el verde heredado de la v1.1 (A, C y F eran las entradas de entonces)
    # se borra ENTERO: la F de hoy es la cotización, que es calculada.
    _vaciar(ws, 5, 4 + motor.CAPACIDAD)
    for f in range(5, 4 + motor.CAPACIDAD + 1):
        ws['C{}'.format(f)].value = None
        ws['D{}'.format(f)].value = 14          # §3 · por defecto 14 pagas
        ws['H{}'.format(f)].value = None
        _formula(ws, 'E{}'.format(f),
                 '=IF(OR($C{f}="",$D{f}=""),"",ROUND($C{f}*$D{f}/12,2))'
                 .format(f=f))
        _formula(ws, 'F{}'.format(f),
                 '=IF($E{f}="","",ROUND($E{f}*$C$2,2))'.format(f=f))
        _formula(ws, 'G{}'.format(f),
                 '=IF($E{f}="","",ROUND($E{f}+$F{f},2))'.format(f=f))
        # RD-20 · el divisor son las semanas EFECTIVAS (F2, verde) partidas
        # entre 12, no las 52 del calendario.
        _formula(ws, 'I{}'.format(f),
                 '=IFERROR(IF(OR($G{f}="",$H{f}="",$H{f}=0,$F$2="",$F$2=0),'
                 '"",ROUND($G{f}/($H{f}*$F$2/12),2)),"")'.format(f=f))

    # 4) TOTALES coherentes con las columnas nuevas
    t = FILA_TOT_03
    ws['A{}'.format(t)] = 'TOTALES'
    ws['A{}'.format(t)].font = Font(bold=True)
    ws['B{}'.format(t)].value = None
    ws['D{}'.format(t)].value = None
    for col in ('C', 'E', 'F', 'G', 'H'):
        _formula(ws, '{}{}'.format(col, t),
                 '=SUM(${c}$5:${c}${u})'.format(c=col, u=t - 1),
                 motor.FMT_EUR if col in ('C', 'E', 'F', 'G')
                 else motor.FMT_DEC2)
    _formula(ws, 'I{}'.format(t),
             '=IFERROR(IF(OR($G${t}="",$H${t}=0,$F$2="",$F$2=0),"",'
             'ROUND($G${t}/($H${t}*$F$2/12),2)),"")'.format(t=t),
             motor.FMT_EUR)
    for col in 'ACDEFGHI':
        ws['{}{}'.format(col, t)].fill = PatternFill('solid', fgColor='ECEFF1')

    _dv(ws, 'D5:D{}'.format(t - 1), ['12', '14', '15'],
        'Número de pagas no válido',
        'Doce pagas = extras prorrateadas mes a mes. Catorce (lo habitual en '
        'hostelería) o quince = las extras se pagan aparte, pero el coste '
        'mensual REAL sigue siendo el prorrateado: es lo que calcula la '
        'columna E.',
        '12, 14 o 15. La columna E prorratea el bruto anual a doce meses.')
    _pie(ws, t + 2, 9)
    cambios.append('03:Nóminas!D4:I{}: pagas prorrateadas, cotización por '
                   'parámetro y coste/hora — la columna «Horas Contratadas» '
                   'deja de estar muerta (DOM-31/TEC-26)'.format(t))
    cambios.append('03:Nóminas!F2: «Semanas efectivas al año» en celda verde '
                   '({} por defecto) y el «Coste/hora» divide por ella en vez '
                   'de por las 52 del calendario: contar como trabajadas las '
                   'semanas de vacaciones y los festivos dejaba el coste por '
                   'hora un 10-13 % corto, y ese número es el que las '
                   'Instrucciones mandan llevar a la tarifa del 02 — RD-20'
                   .format(SEMANAS_EFECTIVAS))


#: RC-11 · el semáforo del 03 sólo conocía SEIS tipos y la FAQ de la landing
#: promete que las plantillas se adaptan a «restaurante casual, fine dining,
#: fast casual, HOTEL, catering, cadenas». Un hotelero caía al 30/35 %
#: genérico mientras el BONUS-02, con sus diez tipos, le daba 42/44 % para el
#: mismo negocio: doce puntos de diferencia en el umbral, y dos ficheros del
#: MISMO kit emitiendo veredictos opuestos sobre el mismo ratio. Lo mismo con
#: pizzería, dark kitchen y heladería. Los cuatro entran con los MISMOS
#: umbrales que `grupo_c.TIPOS_BONUS`.
TIPOS_NEGOCIO_EXTRA = [
    ('Pizzería',                '26-30%', '24-32%', 0.30, 0.32),
    ('Dark Kitchen / Delivery', '22-28%', '20-30%', 0.28, 0.30),
    ('Hotel (restaurante)',     '35-42%', '33-44%', 0.42, 0.44),
    ('Heladería / Obrador',     '25-30%', '22-32%', 0.30, 0.32),
]

#: Geometría de 'Ratio Coste Laboral'. RD-21 mete cuatro filas nuevas en el
#: bloque de coste, así que la tabla de referencia y el pie se calculan a
#: partir de aquí y no se escriben a mano en cinco sitios.
RATIO_TABLA_TIT = 16
RATIO_TABLA_HDR = 17
RATIO_TABLA_R0 = 18


def _n03_ratio(wb, cambios):
    """§3 · DOM-16/TEC-10/TEC-11/COM-11/COM-17 + RD-01/RT-01/RD-21/RC-11.

    Lo que había en la v1.1: `B7='=IF(B4>0,B5/B4*100,0)'` (con la hoja vacía
    daba 0) y `B9='=IF(B7<30,"🟢 EXCELENTE…"…)'`, así que un fichero recién
    descargado abría en verde EXCELENTE; y un fine dining al 38 % —que su
    propia tabla declara correcto— salía en rojo.

    Y lo que quedaba después de la v2.0, medido con pycel sobre la copia
    dry-run:

      · **RD-01/RT-01 · la guarda se puso sólo del lado de las ventas.** `B5`
        era `=Nóminas!$G$35`, un `SUM` sobre una hoja vacía, así que vale 0 y
        no `""`: bastaba teclear las ventas del mes sin haber volcado una sola
        nómina para que el ratio saliera 0,0 % y el semáforo declarara «🟢
        EXCELENTE», con su formato condicional verde. El mismo falso positivo
        tranquilizador del R1, a una pulsación de distancia, en la métrica que
        la landing vende como argumento principal del 03. Ahora la guarda mira
        las DOS entradas, los dos huecos se nombran por separado, y por debajo
        del 60 % del objetivo el veredicto avisa de que el ratio es
        implausible: un coste laboral demasiado BAJO es tan alarma como uno
        alto — significa que faltan nóminas.
      · **RD-21 · el coste laboral era sólo las nóminas, y bloqueado.** No
        había forma de incluir las horas extra del 02 (que el propio kit
        totaliza en `Resumen Mensual!G37` y no lee nadie), ni una ETT, ni el
        refuerzo del día pico que recomienda el BONUS-02. El ratio salía
        sistemáticamente por debajo del real y el semáforo aprobaba de más.
      · **RC-11 · seis tipos de negocio frente a los diez del BONUS-02.**
    """
    ws = wb['Ratio Coste Laboral']
    motor._limpiar_dv(ws)
    tipos = TIPOS_NEGOCIO + TIPOS_NEGOCIO_EXTRA
    r0 = RATIO_TABLA_R0
    r1 = r0 + len(tipos) - 1
    nota = r1 + 1
    pie = nota + 2
    _vaciar(ws, 3, max(pie + 2, ws.max_row))
    _merges(ws, ['A1:E1', 'A2:E2', 'A{0}:E{0}'.format(pie), 'B13:E13'])
    _anchos(ws, {'A': 40, 'B': 20, 'C': 26, 'D': 16, 'E': 16})

    _rotulo(ws, 3, 'Tipo de negocio:')
    _verde(ws, 'B3')
    ws['B3'].value = None
    _dv(ws, 'B3', [t[0] for t in tipos], 'Tipo de negocio no válido',
        'De aquí salen los dos umbrales del semáforo por VLOOKUP sobre la '
        'tabla de esta misma hoja. Son los mismos diez tipos y los mismos '
        'umbrales que la calculadora del BONUS-02: los dos ficheros no pueden '
        'dar veredictos distintos sobre el mismo ratio. Si lo dejas en '
        'blanco, el semáforo usa el 30 % / 35 % genérico.',
        'Elige uno de los diez tipos de la tabla de referencia de abajo.')

    _rotulo(ws, 4, 'Ventas Netas del Mes (€):')
    _verde(ws, 'B4', fmt=motor.FMT_EUR)

    # ---- RD-21 · el coste laboral del MES, no sólo las nóminas -----------
    _rotulo(ws, 5, 'Nóminas del mes (€) — sale de la hoja «Nóminas»:')
    _formula(ws, 'B5', "=Nóminas!$G${}".format(FILA_TOT_03), motor.FMT_EUR)
    for fila, rotulo, prompt in (
            (6, 'Horas extra del mes (€):',
             'Cópialo de «Resumen Mensual»!G37 del fichero 02 (el total de la '
             'columna «Coste de las horas extra»). El kit lo calcula y hasta '
             'ahora no lo leía nadie, así que el ratio salía corto.'),
            (7, 'Personal externo / ETT (€):',
             'Extras de fin de semana, ETT, personal de refuerzo del día pico '
             '(el que recomienda contratar el BONUS-02). Es coste laboral '
             'aunque no esté en tu nómina.'),
            (8, 'Otros costes de personal (€):',
             'Dietas, formación obligatoria, vestuario, transporte, comida de '
             'personal… lo que tu contabilidad impute a personal y no salga '
             'de las nóminas.')):
        _rotulo(ws, fila, rotulo)
        _verde(ws, 'B{}'.format(fila), fmt=motor.FMT_EUR)
        _nota_dv(ws, 'B{}'.format(fila), rotulo.rstrip(':'), prompt)
    _rotulo(ws, 9, 'COSTE LABORAL TOTAL DEL MES (€):', negrita=True)
    _formula(ws, 'B9',
             '=IF(COUNT($B$5:$B$8)=0,"",ROUND(SUM($B$5:$B$8),2))',
             motor.FMT_EUR)
    ws['B9'].font = Font(bold=True)

    # ---- RD-01/RT-01 · el ratio y su semáforo ----------------------------
    _rotulo(ws, 11, 'RATIO COSTE LABORAL (%):', negrita=True)
    _formula(ws, 'B11',
             '=IFERROR(IF(OR($B$4="",$B$4<=0,$B$9="",$B$9<=0),"",'
             'ROUND($B$9/$B$4,4)),"")', motor.FMT_PCT1)
    _rotulo(ws, 11, 'Ratio objetivo de tu tipo:', col=3)
    _formula(ws, 'D11', '=IFERROR(VLOOKUP($B$3,$A${a}:$E${b},4,FALSE),0.3)'
             .format(a=r0, b=r1), motor.FMT_PCT1)
    _rotulo(ws, 12, 'Máximo aceptable de tu tipo:', col=3)
    _formula(ws, 'D12', '=IFERROR(VLOOKUP($B$3,$A${a}:$E${b},5,FALSE),0.35)'
             .format(a=r0, b=r1), motor.FMT_PCT1)

    _rotulo(ws, 13, 'Semáforo:', negrita=True)
    _formula(ws, 'B13',
             '=IF(OR($B$4="",$B$4<=0),'
             '"— introduce las VENTAS netas del mes (B4)",'
             'IF(OR($B$9="",$B$9<=0),'
             '"— aún no hay coste: vuelca las nóminas en la hoja «Nóminas»",'
             'IF($B$11<$D$11*0.6,'
             '"⚠ ratio implausible: revisa que estén TODAS las nóminas",'
             'IF($B$11<$D$11,"🟢 EXCELENTE (por debajo de tu objetivo)",'
             'IF($B$11<=$D$12,"🟡 VIGILAR (en el límite)",'
             '"🔴 ACCIÓN CORRECTIVA")))))', 'General')
    ws['B13'].alignment = Alignment(vertical='center')

    _rotulo(ws, 14, '▸ Sin tipo de negocio elegido el semáforo cae al 30 % / '
                    '35 % genérico. Y por debajo del 60 % de tu objetivo no '
                    'te felicita: un coste laboral así de bajo casi siempre '
                    'significa que faltan nóminas por volcar, no que vayas '
                    'sobrado.')
    ws['A14'].font = Font(size=9, italic=True)
    ws['A14'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[14].height = 28

    _rotulo(ws, RATIO_TABLA_TIT, 'Ratios de Referencia por Tipo', negrita=True)
    _cabecera(ws, RATIO_TABLA_HDR,
              ['Tipo de Negocio', 'Ratio Objetivo', 'Rango Aceptable',
               'Objetivo % (núm.)', 'Aceptable % (núm.)'])
    for i, fila in enumerate(tipos):
        r = r0 + i
        ws['A{}'.format(r)] = fila[0]
        ws['B{}'.format(r)] = fila[1]
        ws['C{}'.format(r)] = fila[2]
        ws['D{}'.format(r)] = fila[3]
        ws['E{}'.format(r)] = fila[4]
        for col in ('B', 'C'):
            ws['{}{}'.format(col, r)].number_format = 'General'
        for col in ('D', 'E'):
            ws['{}{}'.format(col, r)].number_format = motor.FMT_PCT1

    _rotulo(ws, nota, '▸ Las dos columnas numéricas son el extremo superior '
                      'de los rangos de texto: es lo que lee el VLOOKUP del '
                      'semáforo, para que la tabla y el veredicto no puedan '
                      'contradecirse. Son los mismos diez tipos y los mismos '
                      'umbrales que la hoja «Ratios por Tipo» del BONUS-02.')
    ws['A{}'.format(nota)].font = Font(size=9, italic=True)
    ws['A{}'.format(nota)].alignment = Alignment(wrap_text=True,
                                                 vertical='top')
    ws.row_dimensions[nota].height = 28
    _merges(ws, ['A1:E1', 'A2:E2', 'A{0}:E{0}'.format(pie), 'B13:E13',
                 'A14:E14', 'A{0}:E{0}'.format(nota)])
    _pie(ws, pie, 5)
    cambios.append('03:Ratio Coste Laboral!B3/D11/D12/B13: umbrales por '
                   'VLOOKUP sobre A{}:E{} y veredicto en blanco con la hoja '
                   'vacía (DOM-16/TEC-10/TEC-11/COM-17)'.format(r0, r1))
    cambios.append('03:Ratio Coste Laboral!B11/B13: la guarda deja de mirar '
                   'sólo las VENTAS. Antes bastaba teclear las ventas del mes '
                   'sin una sola nómina volcada para que el ratio saliera '
                   '0,0 % y el semáforo dijera «🟢 EXCELENTE» —B5 era un SUM '
                   'sobre una hoja vacía, que vale 0 y no «»—. Ahora los dos '
                   'huecos se nombran por separado y por debajo del 60 % del '
                   'objetivo el veredicto avisa de ratio implausible — '
                   'RD-01/RT-01')
    cambios.append('03:Ratio Coste Laboral!B5:B9: el coste laboral deja de '
                   'ser sólo las nóminas y bloqueado. B5 sigue saliendo de '
                   '«Nóminas» y entran tres verdes —horas extra del mes '
                   '(02!Resumen Mensual!G37), personal externo / ETT y otros '
                   'costes de personal— que B9 suma. El ratio salía '
                   'sistemáticamente por debajo del real y el semáforo '
                   'aprobaba de más — RD-21')
    cambios.append('03:Ratio Coste Laboral!A{}:E{}: la tabla pasa de 6 a {} '
                   'tipos —entran hotel, pizzería, dark kitchen y heladería '
                   'con los MISMOS umbrales que el BONUS-02—, así que un '
                   'hotelero deja de caer al 30/35 % genérico mientras el '
                   'bonus le daba 42/44 % para el mismo negocio — RC-11'
                   .format(r0, r1, len(tipos)))


def _n03_prevision(wb, cambios):
    """§3 · DOM-09/TEC-15/TEC-20/COM-09 — cubiertos por SERVICIO, no por día.

    Lo que había: `B10='=CEILING(B4/B5,1)'` con B4 = 80 cubiertos al DÍA y
    B5 = 20 cubiertos por cocinero, o sea 4 cocineros, 7 camareros y 4 barras
    = **15 personas por turno** para un casual de 80 cubiertos; y un coste
    `=B13*B15*1.30` que multiplicaba esas 15 personas por el salario, sin
    pagas, sin días de apertura y sin factor de cobertura.

    Lo que hay: la MISMA cadena que la SPEC fija para el BONUS-02 —cubiertos
    por servicio → personal por servicio → presencias/día → horas/semana →
    FTE—, que con los valores por defecto da **7 FTE** y 16.292,50 €/mes.
    """
    ws = wb['Previsión por Servicio']
    motor._limpiar_dv(ws)
    # deshacer TODAS las combinaciones antes de vaciar: escribir en una celda
    # de un rango combinado revienta con «'MergedCell' object attribute
    # 'value' is read-only», y aquí hay tres heredadas (A1:D1, A2:D2, A19:D19).
    _merges(ws, [])
    _vaciar(ws)
    _anchos(ws, {'A': 44, 'B': 16, 'C': 14, 'D': 14, 'E': 18, 'F': 4})

    ws['A1'] = 'Previsión de Personal por Servicio'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'AI Chef Pro · aichef.pro — Kit Gestión de Personal y Turnos'
    _seccion(ws, 3, 'ENTRADAS — edita sólo las celdas verdes', 5)

    entradas = [
        (4,  'Cubiertos previstos / día:', 80, motor.FMT_ENT),
        (5,  'Servicios al día (comida + cena):', 2, motor.FMT_ENT),
        (7,  'Cubiertos por cocinero y servicio:', 25, motor.FMT_ENT),
        (8,  'Cubiertos por camarero y servicio:', 22, motor.FMT_ENT),
        (9,  'Cubiertos por persona de barra y servicio:', 80, motor.FMT_ENT),
        (10, 'Horas efectivas por servicio:', 4, motor.FMT_ENT),
        (11, 'Días de apertura / semana:', 6, motor.FMT_ENT),
        (12, 'Jornada contratada (h/semana):', 40, motor.FMT_ENT),
        (13, 'Factor de cobertura (vacaciones, bajas y descansos):', 1.15,
         motor.FMT_DEC2),
        (14, 'Salario medio bruto (€/mes):', 1500, motor.FMT_EUR),
        (15, 'Nº de pagas / año:', 14, motor.FMT_ENT),
        (16, 'Tipo de SS a cargo de la empresa (%):', 0.33, motor.FMT_PCT1),
    ]
    for fila, rotulo, valor, fmt in entradas:
        _rotulo(ws, fila, rotulo)
        _verde(ws, 'B{}'.format(fila), valor, fmt)

    _rotulo(ws, 6, 'Cubiertos por SERVICIO:')
    _formula(ws, 'B6', '=IFERROR(IF(OR($B$4="",$B$5="",$B$5=0),"",'
                       'ROUND($B$4/$B$5,0)),"")', motor.FMT_ENT)
    ws['C6'] = '← es la cifra que dimensiona el turno, no los cubiertos del día'
    ws['C6'].font = Font(size=9, italic=True)

    _dv(ws, 'B5', ['1', '2', '3'], 'Servicios al día no válido',
        'Uno (sólo comidas o sólo cenas), dos (comida y cena) o tres (desayuno, '
        'comida y cena). Es lo que reparte los cubiertos del día entre turnos: '
        'ochenta cubiertos en dos servicios son cuarenta por servicio, y el '
        'equipo se dimensiona con esos cuarenta.',
        'Elige 1, 2 o 3.')

    _seccion(ws, 17, 'RESULTADO — de cubiertos a plantilla', 5)
    for i, txt in enumerate(['Cocina', 'Sala', 'Barra', 'Total / servicio']):
        cel = ws.cell(row=17, column=2 + i, value=txt)
        cel.font = Font(bold=True, size=10)
        cel.alignment = Alignment(horizontal='center')

    _rotulo(ws, 18, 'Personal por SERVICIO:')
    for col, ratio in (('B', '$B$7'), ('C', '$B$8'), ('D', '$B$9')):
        _formula(ws, '{}18'.format(col),
                 '=IFERROR(IF(OR($B$6="",{r}="",{r}=0),0,'
                 'CEILING($B$6/{r},1)),0)'.format(r=ratio), motor.FMT_ENT)
    _formula(ws, 'E18', '=IF($B$6="","",$B$18+$C$18+$D$18)', motor.FMT_ENT)

    _rotulo(ws, 19, 'Presencias al día (personas × servicios):')
    _formula(ws, 'B19', '=IF(OR($E$18="",$B$5=""),"",$E$18*$B$5)',
             motor.FMT_ENT)
    _rotulo(ws, 20, 'Horas de plantilla necesarias / semana:')
    _formula(ws, 'B20', '=IF(OR($B$19="",$B$10="",$B$11=""),"",'
                        '$B$19*$B$10*$B$11)', motor.FMT_ENT)
    _rotulo(ws, 21, 'Plantilla necesaria (FTE):', negrita=True)
    _formula(ws, 'B21', '=IFERROR(IF(OR($B$20="",$B$12="",$B$12=0),"",'
                        'CEILING($B$20/$B$12*$B$13,1)),"")', motor.FMT_ENT)
    _rotulo(ws, 22, 'Coste laboral estimado del equipo (€/mes):', negrita=True)
    _formula(ws, 'B22', '=IFERROR(IF(OR($B$21="",$B$14="",$B$15=""),"",'
                        'ROUND($B$21*$B$14*$B$15/12*(1+$B$16),2)),"")',
             motor.FMT_EUR)

    ws['A24'] = ('▸ FTE = personas a jornada completa, no personas por turno: '
                 'las 240 h/semana que salen por defecto son siete contratos '
                 'de 40 h con el 15 % de cobertura de libranzas, vacaciones y '
                 'bajas ya dentro.')
    ws['A25'] = ('▸ El coste es bruto de convenio × las pagas que elijas, '
                 'prorrateado a doce meses y con la cotización empresarial '
                 'encima. Para el coste REAL de tu equipo usa la hoja '
                 '«Nóminas», que parte de tus contratos.')
    for r in (24, 25):
        ws.cell(row=r, column=1).font = Font(size=9, italic=True)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True,
                                                       vertical='top')
        ws.row_dimensions[r].height = 28
    _pie(ws, 26, 5)
    _merges(ws, ['A1:E1', 'A2:E2', 'A24:E24', 'A25:E25', 'A26:E26'])
    cambios.append('03:Previsión por Servicio!B6/B18:E18/B21/B22: un solo '
                   'modelo de dimensionamiento (cubiertos por SERVICIO → '
                   '7 FTE → 16.292,50 €/mes) — antes 15 personas por turno '
                   '(DOM-09/TEC-15/TEC-20/COM-09)')


def _n03_instrucciones(wb, cambios):
    ws = wb['Instrucciones']
    _instrucciones(ws, [
        '03 · Coste Laboral Mensual',
        None,
        'Cómo usar esta plantilla:',
        "▸ Vuelca los contratos en la hoja 'Nóminas': bruto mensual, número de "
        'pagas al año y horas contratadas por semana. Lo demás se calcula.',
        '▸ El tipo de cotización empresarial vive en UNA celda verde (C2 de '
        "'Nóminas'), no dentro de las fórmulas: ajústalo a tu CNAE, a tu tipo "
        'de contrato y a tu convenio. El 33 % por defecto es lo que ronda un '
        'indefinido general sumando contingencias comunes, desempleo, AT/EP, '
        'FOGASA, formación profesional y MEI.',
        '▸ Las pagas extra se PRORRATEAN: catorce pagas de 1.500 € no cuestan '
        '1.500 € al mes, cuestan 1.750 €. La columna «Bruto prorrateado/mes» '
        'es la que hay que mirar para presupuestar.',
        '▸ La columna «Coste/hora» es el número que hay que llevar a la '
        "tarifa horaria del fichero 02: es coste real de empresa, no el bruto "
        'del convenio.',
        "▸ La hoja 'Ratio Coste Laboral' compara ese coste con tus ventas "
        'netas del mes y lo juzga con los umbrales de TU tipo de negocio, '
        'elegido en la celda verde B3.',
        "▸ La hoja 'Previsión por Servicio' dimensiona el equipo al revés: de "
        'cubiertos por servicio a plantilla necesaria en jornadas completas.',
        None,
        'Lo que cambia respecto de la versión 1.1:',
        '▸ El semáforo ya no da «EXCELENTE» con la hoja en blanco: sin ventas '
        'introducidas no hay veredicto.',
        '▸ Y ya no suspende a la alta cocina: un fine dining al 38 % está '
        'donde debe estar, y un fast casual al 38 % está en pérdidas. Los dos '
        'umbrales salen de la tabla de referencia de esa misma hoja.',
        '▸ La previsión de personal razona por SERVICIO. Ochenta cubiertos al '
        'día en dos servicios son cuarenta por servicio: dimensionar el turno '
        'con los ochenta duplicaba la plantilla.',
        None,
        'Aviso: esto es una herramienta de gestión, no un cálculo de nómina. '
        'La cotización real depende de tu CNAE, del grupo de cotización, de '
        'las bonificaciones que te apliquen y del convenio. Contrástalo con '
        'tu gestoría antes de firmar nada.',
    ])
    cambios.append('03:Instrucciones!B2:B22: reescritas — desaparece «La '
                   'Seguridad Social empresa (~30%)» y el semáforo fijo de '
                   '30/35 % (DOM-08/COM-17)')


# ==========================================================================
# 04 · ONBOARDING NUEVO EMPLEADO  (SPEC §4, bloque 04)
# ==========================================================================
#: Las TRES tareas que faltaban y tienen plazo legal (DOM-15/§7-bis.3). Entran
#: al final del primer tramo, que pasa de 10 a 13 filas y con él el checklist
#: de 47 a **50** tareas. `(texto, plazo)`.
TAREAS_NUEVAS_04 = [
    ('Comunicación del contrato a Contrat@ (SEPE) — 10 días hábiles desde la '
     'firma', 'Día 7'),
    ('Copia básica del contrato a la representación legal de los trabajadores '
     '(RLT) — 10 días', 'Día 7'),
    ('Modelo 145 de IRPF firmado (situación personal y familiar para calcular '
     'la retención)', 'Día 1'),
]

#: DOM-15 · el alta en la Seguridad Social no es «papeleo del primer día»:
#: tiene que estar hecha ANTES de que el trabajador empiece a trabajar.
TAREA_ALTA_SS = ('Alta en Seguridad Social (TA.2/S) — OBLIGATORIO ANTES del '
                 'inicio de la jornada (art. 32.3 RD 84/1996)')

#: DOM-22/TEC-12 · la columna «Categoría» venía cortada a 14 caracteres
#: exactos («FORMACIÓN OBLI», «PERIODO DE PRU»): un `slice[:14]` del generador.
CATEGORIAS_04 = ['DOCUMENTACIÓN LEGAL', 'FORMACIÓN OBLIGATORIA',
                 'EQUIPAMIENTO Y ACCESOS', 'FORMACIÓN OPERATIVA',
                 'PERIODO DE PRUEBA']

#: §4 · «Fecha Límite» estaba vacía en las 47 filas. Se precarga con el plazo
#: real de cada tarea, en la escala del propio checklist: Día −1 (antes del
#: alta), Día 1, Día 7 y Día 30. El signo menos es un guion normal a propósito:
#: el U+2212 degenera al pasar por un heredoc del shell.
PLAZOS_04 = [
    # 1 · Documentación legal y administrativa (13)
    ['Día -1', 'Día -1', 'Día -1', 'Día -1', 'Día 1', 'Día 1', 'Día 1',
     'Día 1', 'Día 1', 'Día 7', 'Día 7', 'Día 7', 'Día 1'],
    # 2 · Formación obligatoria (8)
    ['Día 1', 'Día 7', 'Día 7', 'Día 1', 'Día 7', 'Día 7', 'Día 1', 'Día 30'],
    # 3 · Equipamiento y accesos (9)
    ['Día 1', 'Día 1', 'Día 1', 'Día 1', 'Día 7', 'Día 1', 'Día 1', 'Día 7',
     'Día 1'],
    # 4 · Formación operativa (12)
    ['Día 1', 'Día 1', 'Día 1', 'Día 1', 'Día 7', 'Día 7', 'Día 7', 'Día 7',
     'Día 7', 'Día 7', 'Día 30', 'Día 30'],
    # 5 · Periodo de prueba y evaluación (8)
    ['Día 1', 'Día 7', 'Día 30', 'Día 7', 'Día 30', 'Día 30', 'Día 30',
     'Día 30'],
]


def _n04(wb, cambios):
    """§4 · DOM-01/DOM-15/DOM-22/DOM-28/TEC-02/TEC-12/TEC-13/COM-03.

    El defecto estrella: `C68='=COUNTIF(F7:F65,"✓")'` barre un rango CONTINUO
    que se traga las cabeceras de las secciones 2ª a 5ª —`F19`, `F30`, `F42` y
    `F57` valen literalmente «✓»—, así que un checklist en blanco abre en
    «4 de 51 completadas» y un 8,51 %. Y es justo la mejora que el changelog
    v1.1 vende como estrella.

    Se hacen LAS DOS cosas que apunta la SPEC: las cabeceras pasan a decir
    «Hecho» (para que nunca vuelva a haber un «✓» que no sea una tarea hecha)
    **y** los contadores se acotan por tramo. Con una sola de las dos, el
    siguiente que añada una sección volvería a romperlo.
    """
    ws = wb['Checklist Onboarding']
    secciones = motor.secciones_04(ws)

    # 1) las tres tareas que faltaban (DOM-15). Centinela de idempotencia: el
    #    primer tramo mide 10 filas en la v1.1 y 13 cuando ya están puestas.
    r_tit, hdr, r0, r1 = secciones[0]
    if r1 - r0 + 1 == 10:
        for _ in TAREAS_NUEVAS_04:
            motor.insertar_fila(ws, r1 + 1)
        for i, par in enumerate(TAREAS_NUEVAS_04):
            f = r1 + 1 + i
            ws.cell(row=f, column=1, value=r1 - r0 + 2 + i)
            ws.cell(row=f, column=2, value=par[0])
        cambios.append('04:Checklist Onboarding!B{}:B{}: +3 tareas con plazo '
                       'legal (Contrat@/SEPE, copia básica a la RLT, modelo '
                       '145) — 47 → 50 (DOM-15)'.format(r1 + 1, r1 + 3))
        secciones = motor.secciones_04(ws)

    # 2) alta en la SS: ANTES del inicio de la jornada
    for f in range(secciones[0][2], secciones[0][3] + 1):
        v = ws.cell(row=f, column=2).value
        if isinstance(v, str) and v.startswith('Alta en Seguridad Social'):
            if v != TAREA_ALTA_SS:
                ws.cell(row=f, column=2).value = TAREA_ALTA_SS
                cambios.append('04:Checklist Onboarding!B{}: el alta en la SS '
                               'pasa a ser previa al inicio de la jornada '
                               '(art. 32.3 RD 84/1996, DOM-15)'.format(f))

    # 3) cabeceras «✓» → «Hecho», categorías completas y plazos
    tramos = []
    for i, sec in enumerate(secciones):
        r_tit, hdr, r0, r1 = sec
        tramos.append((r0, r1))
        ws.cell(row=hdr, column=6).value = 'Hecho'
        ws.cell(row=hdr, column=5).value = 'Fecha Límite'
        plazos = PLAZOS_04[i] if i < len(PLAZOS_04) else []
        for j, f in enumerate(range(r0, r1 + 1)):
            ws.cell(row=f, column=1).value = j + 1
            ws.cell(row=f, column=3).value = CATEGORIAS_04[i]
            if j < len(plazos):
                ws.cell(row=f, column=5).value = plazos[j]
    _anchos(ws, {'A': 5, 'B': 52, 'C': 22, 'D': 18, 'E': 14, 'F': 10,
                 'G': 24})

    # ⚠ `motor.insertar_fila` mueve valores, estilos, combinaciones y DV, pero
    # NO el formato condicional: la única regla que trae la v1.1 —«pinta la
    # fila entera de verde cuando la tarea está hecha», `A7:G65` con
    # `$F7="✓"`— se quedaba cubriendo hasta la fila 65 con las tareas llegando
    # a la 68, así que las tres últimas del periodo de prueba nunca se
    # pintaban. Se reengancha al rango real conservando su propio `dxf`.
    ultima = tramos[-1][1]
    viejas, refs = [], set()
    for bloque in list(ws.conditional_formatting):
        if str(bloque.sqref).startswith('A7:G'):
            viejas.extend(bloque.rules)
            refs.add(str(bloque.sqref))
    if viejas:
        motor._limpiar_cf(ws, refs)
        for regla in viejas:
            ws.conditional_formatting.add('A7:G{}'.format(ultima), regla)
        cambios.append('04:Checklist Onboarding!A7:G{}: la regla de formato '
                       'condicional heredada llegaba sólo a la fila 65 — '
                       '`insertar_fila` no desplaza el formato condicional'
                       .format(ultima))

    # 4) validación de «Hecho» sobre TODOS los tramos, incluidas las 3 nuevas.
    #    Se retiran todas las DV de la hoja antes: la heredada de la v1.1
    #    enumeraba las 47 celdas una a una y no cubría las filas nuevas.
    ws.data_validations.dataValidation = []
    dv = DataValidation(
        type='list', formula1='"✓,✗,—"', allow_blank=True,
        showErrorMessage=True, errorTitle='Marca no válida',
        error='✓ hecha · ✗ pendiente · — no aplica a este puesto.',
        errorStyle='stop', showInputMessage=True,
        promptTitle='{} · Hecho'.format(MARCA_B),
        prompt='Marca «—» en lo que NO aplique a este puesto (un cocinero de '
               'partida no necesita el acceso al TPV): esas tareas salen del '
               'denominador y el progreso puede llegar al 100 %.')
    ws.add_data_validation(dv)
    for r0, r1 in tramos:
        dv.add('F{}:F{}'.format(r0, r1))

    # 5) los contadores, acotados por TRAMO
    fila_res = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().upper() == 'RESUMEN':
            fila_res = r
            break
    if fila_res is None:
        return
    hechas = '+'.join('COUNTIF($F${}:$F${},"✓")'.format(a, b)
                      for a, b in tramos)
    total = '+'.join('COUNTIF($B${}:$B${},"?*")'.format(a, b)
                     for a, b in tramos)
    noaplica = '+'.join('COUNTIF($F${}:$F${},"—")'.format(a, b)
                        for a, b in tramos)
    f1, f2, f3 = fila_res + 1, fila_res + 2, fila_res + 3
    ws.cell(row=f1, column=2, value='Tareas completadas:')
    _formula(ws, 'C{}'.format(f1), '=' + hechas, motor.FMT_ENT)
    ws.cell(row=f1, column=4, value='de')
    _formula(ws, 'E{}'.format(f1), '=' + total, motor.FMT_ENT)
    ws.cell(row=f2, column=2,
            value='Total tareas aplicables (sin las marcadas «—»):')
    _formula(ws, 'C{}'.format(f2),
             '=$E${}-({})'.format(f1, noaplica), motor.FMT_ENT)
    ws.cell(row=f3, column=2, value='Progreso:')
    _formula(ws, 'C{}'.format(f3),
             '=IF($C${d}>0,$C${n}/$C${d},0)'.format(n=f1, d=f2),
             motor.FMT_PCT1)
    for f in (f1, f2, f3):
        ws.cell(row=f, column=2).font = Font(bold=True)
    ws.cell(row=fila_res, column=1).font = Font(bold=True, size=11)
    cambios.append('04:Checklist Onboarding!C{}/C{}/C{}: contadores acotados '
                   'por tramo y denominador sin las tareas «—» — el checklist '
                   'recién descargado marca 0 %, no 8,51 % '
                   '(DOM-01/DOM-28/TEC-02/TEC-13)'.format(f1, f2, f3))


def _n04_instrucciones(wb, cambios):
    ws = wb['Instrucciones']
    _instrucciones(ws, [
        '04 · Onboarding Nuevo Empleado',
        None,
        'Cómo usar esta plantilla:',
        '▸ Duplica el fichero para cada incorporación y escribe arriba el '
        'nombre, el puesto y la fecha de alta.',
        '▸ Marca cada tarea en la columna «Hecho»: ✓ hecha · ✗ pendiente · '
        '— no aplica a este puesto.',
        '▸ La «—» es importante: saca la tarea del denominador. Un cocinero '
        'de partida que no necesita el acceso al TPV puede llegar al 100 % '
        'igual que un jefe de sala.',
        '▸ La columna «Fecha Límite» viene precargada con el plazo de cada '
        'tarea en días desde el alta: Día -1 es ANTES de que empiece a '
        'trabajar. Sustitúyela por la fecha real si prefieres el calendario.',
        '▸ El resumen del final cuenta tramo a tramo, así que las cabeceras '
        'de sección ya no se cuelan como tareas hechas: un checklist en '
        'blanco marca 0 %.',
        None,
        'Categorías (cada sección lleva su color):',
        '▸ Azul = Documentación legal y administrativa',
        '▸ Naranja = Formación obligatoria',
        '▸ Verde = Equipamiento y accesos',
        '▸ Morado = Formación operativa',
        '▸ Rojo = Periodo de prueba y evaluación',
        None,
        'Plazos que no dependen de ti:',
        '▸ El alta en la Seguridad Social tiene que estar presentada ANTES '
        'del inicio de la jornada (art. 32.3 del RD 84/1996). No es papeleo '
        'del primer día: es requisito para que ese primer día sea legal.',
        '▸ El contrato se comunica a Contrat@ (SEPE) dentro de los 10 días '
        'hábiles siguientes a su formalización, y la copia básica va a la '
        'representación legal de los trabajadores en 10 días.',
        '▸ El modelo 145 lo firma la persona para que puedas calcular bien '
        'su retención de IRPF; sin él, la retención se aplica sin cargas '
        'familiares.',
        None,
        'Aviso: los plazos son los del Estatuto de los Trabajadores y su '
        'normativa de desarrollo a agosto de 2026. Tu convenio puede exigir '
        'más trámites, nunca menos.',
    ])
    cambios.append('04:Instrucciones!B2:B26: reescritas — «Hecho», el papel '
                   'de la marca «—» y los tres plazos legales nuevos')


# ==========================================================================
# 05 · PLANIFICACIÓN DE VACACIONES  (SPEC §4, bloque 05)
# ==========================================================================
SEMANAS = 53
COL_S1 = 'B'                                    # semana 1
COL_SN = get_column_letter(1 + SEMANAS)         # BB · semana 53
COL_TOT = get_column_letter(2 + SEMANAS)        # BC · «Días Usados»
EMP0, EMP1 = 6, 5 + motor.CAPACIDAD             # 6..35
#: RD-12/RD-13/RT-11 · la fila de ausencias se DESDOBLA. 36 temporada ·
#: 37 lo marcado a mano en la rejilla · 38 lo que sale de las solicitudes
#: aprobadas · 39 la mayor de las dos · 40 la alerta.
F_TEMP = EMP1 + 1                               # 36
F_AUS = EMP1 + 2                                # 37 · rejilla
F_SOL = EMP1 + 3                                # 38 · solicitudes aprobadas
F_MAX = EMP1 + 4                                # 39 · la mayor
F_ALERTA = EMP1 + 5                             # 40

#: RD-09/RT-03 · «Solicitudes» NO está indexada por empleado, está indexada
#: por PETICIÓN, y estaba dimensionada a 30 filas «porque §1.3 dice 30
#: empleados». Con 30 personas eso es UNA sola solicitud por persona y año:
#: nadie coge los 30 días del tirón —en hostelería se parten en verano,
#: Navidad y algún puente—, así que la hoja se agotaba a los 10-15 empleados
#: y a partir de ahí el saldo dejaba de ser cierto EN SILENCIO (los SUMIFS
#: del saldo y el INDEX/MATCH de cobertura tenían el rango cerrado en la 34,
#: y una fila 35 escrita a mano ni sumaba ni traía fórmulas). El 02 ya subió
#: a 300 filas por este mismo razonamiento (COM-18).
SOL0 = 5
SOL_FILAS = 150                                 # 30 personas × 5 periodos
SOL1 = SOL0 + SOL_FILAS - 1                     # 154

#: RD-11/RT-13/RC-17 · el calendario estaba anclado a 2027 con DOS literales
#: (`datetime(2027, 1, 4)` y `datetime(2027, 12, 31)`) en un producto sellado
#: «Versión 2.0 · agosto 2026»: quien compraba en 2026 no encontraba el año
#: que estaba gestionando, y quien compre en 2029 seguiría recibiendo 2027 —
#: con el agravante de que esa fecha de cierre es el input del prorrateo
#: (RT-08). Ahora se DERIVA de la fecha de construcción: el primer lunes del
#: año siguiente (el calendario de vacaciones se cierra con dos meses de
#: antelación, art. 38.3 ET). Sigue siendo una celda VERDE y `main.py` tiene
#: un gate que falla si el año del fichero es anterior al año en curso.
def _primer_lunes(anio):
    d = datetime.datetime(anio, 1, 1)
    return d + datetime.timedelta(days=(7 - d.weekday()) % 7)


ANIO_CALENDARIO = datetime.datetime.now().year + 1
LUNES_SEMANA_1 = _primer_lunes(ANIO_CALENDARIO)

#: Temporada alta por defecto en un negocio de hostelería español: julio y
#: agosto y la Navidad. RT-13 · se calcula por el MES del lunes de cada
#: semana, no por su NÚMERO: el número de semana no se mueve si cambia el
#: ancla, y la banda gris del formato condicional —que sí usa `MONTH(B$5)`—
#: se descuadraba con ella. Se precarga y se edita.
def _temporada(lunes):
    if lunes.month in (7, 8):
        return 'Alta'
    if lunes.month == 12 and lunes.day >= 15:
        return 'Alta'
    return 'Normal'

#: Puestos reales de una plantilla de restauración, para la DV de la tabla de
#: sustituciones. No es una lista genérica de RR. HH.: son las categorías con
#: las que se cubre un turno.
PUESTOS = ['Jefe/a de cocina', 'Segundo/a de cocina', 'Cocinero/a',
           'Ayudante de cocina', 'Pastelero/a', 'Office',
           'Jefe/a de sala', 'Camarero/a', 'Ayudante de camarero/a',
           'Barman / Bartender', 'Sumiller', 'Host / Recepción',
           'Repartidor/a', 'Encargado/a']


def _rehacer(wb, titulo, indice):
    """Reconstruye una hoja entera en su posición. Es la forma más barata de
    ser idempotente cuando cambia la GEOMETRÍA: el calendario pasa de 12
    columnas (una por MES) a 55 (una por semana + el total), y parchear eso
    dejaría restos de la rejilla vieja a la derecha."""
    if titulo in wb.sheetnames:
        wb.remove(wb[titulo])
    return wb.create_sheet(titulo, indice)


#: ⚠ El `&""` de los criterios de `SUMIFS`/`COUNTIF` NO es decorativo: el
#: `IF` de pycel NO es perezoso —evalúa las DOS ramas— y con la celda del
#: nombre VACÍA el criterio llega como `None`, que su `criteria_parser`
#: rechaza («Couldn't parse criteria: None»). Sin el `&""`, 60 fórmulas del
#: 05 se quedaban sin valor cacheado y las demostraciones leían
#: `ERR:FormulaEvalError` donde tiene que haber una celda en blanco. Medido
#: el 2026-08-24. En Excel es inocuo: la rama nunca se ejecuta.
#: RD-10 · los SUMIFS del saldo sumaban las solicitudes SIN filtrar por año.
#: En cuanto el libro se reutiliza el año siguiente —que es lo que invita a
#: hacer un calendario «Anual»— las vacaciones de 2028 y las de 2029
#: descuentan del mismo derecho de 30 días y todo el mundo aparece con saldo
#: negativo. Se añaden los dos criterios de fecha contra la fecha de INICIO,
#: verificados con pycel (SUMIFS con cuatro criterios evalúa).
def _sumifs_solicitudes(celda_nombre, estado):
    return ('SUMIFS(Solicitudes!$D${a}:$D${b},'
            'Solicitudes!$A${a}:$A${b},{n}&"",'
            'Solicitudes!$G${a}:$G${b},"{e}",'
            'Solicitudes!$B${a}:$B${b},">="&\'Saldo Vacaciones\'!$E$2,'
            'Solicitudes!$B${a}:$B${b},"<="&\'Saldo Vacaciones\'!$G$2)'
            .format(a=SOL0, b=SOL1, n=celda_nombre, e=estado))


def _n05_calendario(wb, cambios):
    """§4 · DOM-04/TEC-04/COM-06 — el calendario cuenta DÍAS.

    Lo que había: **una celda por MES** (`B5:M5` = Ene..Dic) y
    `N6='=IF(B6="V",1,0)+…'`, doce sumandos: quien disfruta agosto entero
    figuraba con **1 día usado y 29 restantes**. Y el 30 del convenio estaba
    escrito dentro de 30 fórmulas de `Solicitudes!F` (COM-24).

    Lo que hay: una fila por empleado × 53 semanas, y el cómputo NO sale de
    las celdas pintadas sino de las FECHAS de las solicitudes aprobadas —que
    es lo que elimina la doble fuente de verdad—.
    """
    ws = _rehacer(wb, 'Calendario Anual', 1)

    # RC-17 · el título dejaba el año en blanco («— Año: ______»), así que
    # nada en pantalla avisaba de para qué ejercicio venía preparada la hoja:
    # había que darse cuenta leyendo la fecha del lunes de la semana 1.
    _formula(ws, 'A1',
             '="Calendario de Vacaciones — Año "&TEXT(YEAR($B$5),"0")')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'AI Chef Pro · aichef.pro — Kit Gestión de Personal y Turnos'
    ws['A3'] = 'Códigos:'
    ws['A3'].font = Font(bold=True, size=9)

    # fila 4 · número de semana · fila 5 · lunes de esa semana
    ws['A4'] = 'Semana nº'
    ws['A5'] = 'Lunes de la semana'
    for c in (1,):
        ws.cell(row=4, column=c).font = Font(bold=True, size=9)
        ws.cell(row=5, column=c).font = Font(bold=True, size=9)
    for n in range(1, SEMANAS + 1):
        col = get_column_letter(1 + n)
        cel = ws['{}4'.format(col)]
        cel.value = n
        cel.font = Font(bold=True, size=8, color='FFFFFF')
        cel.fill = PatternFill('solid', fgColor=motor.CAB)
        cel.alignment = Alignment(horizontal='center')
        fecha = ws['{}5'.format(col)]
        if n == 1:
            fecha.value = LUNES_SEMANA_1
        else:
            anterior = get_column_letter(n)
            _formula(ws, '{}5'.format(col), '={}5+7'.format(anterior))
        fecha.number_format = 'dd/mm'
        fecha.font = Font(size=8)
        fecha.alignment = Alignment(horizontal='center')
        ws.column_dimensions[col].width = 2.4
    motor.marcar_verde(ws, 'B5')
    # RD-11 · B5 es EL parámetro del que cuelga la hoja entera y no estaba
    # rotulado como tal. El SPEC §4 pedía además dos verdes en esta hoja —«B2
    # Lunes de la semana 1» y «B3 Días de vacaciones por convenio»— y ninguna
    # existía: B2 estaba vacía y B3 llevaba la leyenda.
    _dv(ws, 'B5', None, 'Fecha no válida',
        'Lunes de la SEMANA 1 — cámbialo cada año y las 52 fechas restantes '
        'se recalculan solas (cada una es la anterior + 7). De aquí salen '
        'también el año del título, la fecha de cierre del saldo y la banda '
        'gris que separa los meses.',
        'Escribe una fecha posterior al año 2000 (el lunes con el que '
        'quieres empezar el año).',
        marca='Ancla', tipo='custom', formula='B5>36526')   # 01/01/2000
    ws['A5'] = 'Lunes de la semana 1 →'
    ws['A5'].font = Font(bold=True, size=9)
    ws['A5'].alignment = Alignment(horizontal='right')

    ws['{}5'.format(COL_TOT)] = 'Días Usados'
    ws['{}5'.format(COL_TOT)].font = Font(bold=True, size=9)
    ws['{}5'.format(COL_TOT)].alignment = Alignment(wrap_text=True,
                                                    horizontal='center')

    # 30 empleados × 53 semanas
    for f in range(EMP0, EMP1 + 1):
        _formula(ws, '{}{}'.format(COL_TOT, f),
                 '=IF($A{f}="","",{s})'.format(f=f, s=_sumifs_solicitudes(
                     '$A{}'.format(f), 'Aprobado')))
    motor.marcar_verde(ws, 'A{}:{}{}'.format(EMP0, COL_SN, EMP1))
    for f in range(EMP0, EMP1 + 1):
        for n in range(1, SEMANAS + 1):
            ws.cell(row=f, column=1 + n).alignment = Alignment(
                horizontal='center')
            ws.cell(row=f, column=1 + n).font = Font(size=8)

    # fila 36 · temporada · fila 37 · ausencias · fila 38 · alerta
    ws['A{}'.format(F_TEMP)] = 'Temporada (Alta / Normal)'
    ws['A{}'.format(F_AUS)] = 'Ausencias marcadas en la rejilla (V, B, PE)'
    ws['A{}'.format(F_SOL)] = 'Ausencias por solicitudes APROBADAS'
    ws['A{}'.format(F_MAX)] = 'Ausencias en la semana (la mayor de las dos)'
    ws['A{}'.format(F_ALERTA)] = 'Cobertura (alerta)'
    for f in (F_TEMP, F_AUS, F_SOL, F_MAX, F_ALERTA):
        ws.cell(row=f, column=1).font = Font(bold=True, size=9)
    for n in range(1, SEMANAS + 1):
        col = get_column_letter(1 + n)
        lunes = LUNES_SEMANA_1 + datetime.timedelta(days=7 * (n - 1))
        ws['{}{}'.format(col, F_TEMP)] = _temporada(lunes)
        ws['{}{}'.format(col, F_TEMP)].font = Font(size=7)
        ws['{}{}'.format(col, F_TEMP)].alignment = Alignment(
            horizontal='center', text_rotation=90)
        # RD-13 · la fila contaba SÓLO V y B. Un permiso retribuido (PE) es
        # una ausencia que hay que cubrir igual —boda, mudanza, fallecimiento,
        # examen—, está en la leyenda única, está en la DV de la propia
        # rejilla y el formato condicional lo pinta de naranja: la semana con
        # dos personas de permiso se veía como una semana sin ausencias. El
        # festivo (F) NO suma: no es una ausencia individual.
        _formula(ws, '{}{}'.format(col, F_AUS),
                 '=' + '+'.join('COUNTIF({c}${a}:{c}${b},"{q}")'
                                .format(c=col, a=EMP0, b=EMP1, q=q)
                                for q in ('V', 'B', 'PE')))
        # RD-12 · la doble fuente de verdad seguía viva, sólo que repartida:
        # el SALDO salía de las solicitudes y la COBERTURA de las 1.590 celdas
        # pintadas a mano. Quien trabaja con «Solicitudes» —que es lo que
        # dicen las Instrucciones— dejaba esta rejilla en blanco y la promesa
        # de cobertura no se cumplía un lunes. Esta fila lee las solicitudes
        # aprobadas cuyo periodo SOLAPA con la semana.
        _formula(ws, '{}{}'.format(col, F_SOL),
                 '=SUMPRODUCT((Solicitudes!$A${a}:$A${b}<>"")*'
                 '(Solicitudes!$G${a}:$G${b}="Aprobado")*'
                 '(Solicitudes!$B${a}:$B${b}<={c}$5+6)*'
                 '(Solicitudes!$C${a}:$C${b}>={c}$5))'
                 .format(a=SOL0, b=SOL1, c=col))
        _formula(ws, '{}{}'.format(col, F_MAX),
                 '=MAX({c}${s},{c}${p})'.format(c=col, s=F_AUS, p=F_SOL))
        # RT-11 · la alerta mezclaba magnitud y temporada, y la más grave
        # quedaba tapada por la más frecuente: la rama de temporada alta se
        # evaluaba PRIMERO y con `>0`, así que UNA sola ausencia en cualquiera
        # de las 11 semanas precargadas como «Alta» pintaba el bloqueante ⛔ —
        # y con 10 ausencias sobre un máximo de 4 el «⚠ EXCESO», la única
        # alerta que compara contra el parámetro que el cliente ha fijado, no
        # aparecía NUNCA. Ahora manda la magnitud y la temporada agrava, con
        # su propio umbral en celda (Cobertura!D3).
        _formula(ws, '{}{}'.format(col, F_ALERTA),
                 '=IF({c}${m}=0,"",'
                 'IF({c}${m}>Cobertura!$B$3,'
                 'IF({c}${t}="Alta","⛔ EXCESO en TEMP. ALTA","⚠ EXCESO"),'
                 'IF(AND({c}${t}="Alta",{c}${m}>Cobertura!$D$3),'
                 '"⛔ TEMP. ALTA","")))'
                 .format(c=col, m=F_MAX, t=F_TEMP))
        for f in (F_AUS, F_SOL, F_MAX):
            ws['{}{}'.format(col, f)].font = Font(size=8)
            ws['{}{}'.format(col, f)].alignment = Alignment(
                horizontal='center')
        ws['{}{}'.format(col, F_ALERTA)].font = Font(size=7)
    motor.marcar_verde(ws, 'B{f}:{c}{f}'.format(f=F_TEMP, c=COL_SN))
    ws.row_dimensions[F_TEMP].height = 42

    # ⚠ La fila de temporada NO puede llevar una DV de LISTA. Dos gates del
    # motor lo impiden, y los dos son correctos: `leyenda_coherente` trata
    # cualquier lista de valores de 1-2 caracteres de ESTA hoja como códigos
    # de la leyenda única (adiós a «S»/«N»), y `main.demo_leyenda` exige que
    # la hoja tenga EXACTAMENTE UNA lista inline —la de los códigos de
    # ausencia—, así que hasta «Alta,Normal» la tumbaba. Se resuelve con una
    # validación `custom`: sin desplegable, pero con mensaje de entrada y
    # rechazo de lo que no sea Alta/Normal. Las 53 celdas van precargadas
    # justo por eso: el cliente sólo cambia texto que ya está escrito.
    # Su `promptTitle` NO lleva la marca del motor a propósito: `aplicar_dv`
    # limpia por marca TODAS las DV de esta hoja antes de escribir la suya.
    _dv(ws, 'B{f}:{c}{f}'.format(f=F_TEMP, c=COL_SN), None,
        'Temporada no válida',
        'Escribe «Alta» en tus semanas de máxima demanda. Una ausencia en una '
        'semana de temporada alta salta en rojo aunque no supere el máximo de '
        'ausencias simultáneas.', 'Sólo se admite Alta o Normal.',
        marca='Temporada', tipo='custom',
        formula='OR(B{f}="Alta",B{f}="Normal")'.format(f=F_TEMP))

    _anchos(ws, {'A': 22, COL_TOT: 11})
    ws['A{}'.format(F_ALERTA + 2)] = motor.PIE
    ws['A{}'.format(F_ALERTA + 2)].font = Font(size=8, color='888888')
    _merges(ws, ['A1:{}1'.format(COL_TOT), 'A2:{}2'.format(COL_TOT),
                 'A{0}:{1}{0}'.format(F_ALERTA + 2, COL_TOT)])
    cambios.append('05:Calendario Anual!B4:{}{}: rejilla de 12 meses → 53 '
                   'semanas; «Días Usados» ({}) suma los DÍAS de las '
                   'solicitudes aprobadas en vez de contar meses marcados '
                   '(DOM-04/TEC-04/COM-06)'
                   .format(COL_SN, F_ALERTA, COL_TOT))


def _n05_saldo(wb, cambios):
    """§4 · DOM-19/TEC-18/COM-24 — hoja nueva `Saldo Vacaciones`.

    El derecho deja de estar escrito en 30 fórmulas (`Solicitudes!F5:F34` era
    `=IF(E5<>"",30-E5,"")`) y pasa a UNA celda verde, con prorrateo para quien
    entra a mitad de año y con el saldo descontando ya la solicitud EN CURSO,
    que es lo que impide aprobar dos veces los mismos días.
    """
    ws = _rehacer(wb, 'Saldo Vacaciones', 3)
    ws['A1'] = 'Saldo de Vacaciones por Empleado'
    ws['A1'].font = Font(bold=True, size=14)

    coord = motor.parametro(ws, 2, 'dias_convenio', col_rotulo=1, col_valor=2)
    # RD-11/RT-13 · las dos fechas del ejercicio DERIVAN del ancla del
    # calendario en vez de ser dos literales de 2027 congelados en el
    # generador. Y RD-10 necesitaba una fecha de INICIO que no existía: sin
    # ella, los SUMIFS del saldo sumaban las solicitudes de todos los años.
    ws['D2'] = 'Inicio del año:'
    _formula(ws, 'E2', '=DATE(YEAR($G$2),1,1)', motor.FMT_FECHA)
    ws['F2'] = 'Fecha de cierre del año:'
    _formula(ws, 'G2',
             "=DATE(YEAR('Calendario Anual'!$B$5),12,31)", motor.FMT_FECHA)
    for coord_rot in ('D2', 'F2'):
        ws[coord_rot].font = Font(bold=True)
    ws['H2'] = ('← salen del lunes de la semana 1 (Calendario Anual!B5): '
                'cámbialo allí y el ejercicio entero se mueve')
    ws['H2'].font = Font(size=9, italic=True, color='666666')
    ws['A3'] = ('▸ Los nombres salen del Calendario Anual. Rellena la fecha '
                'de alta (para prorratear a quien entra a mitad de año), la '
                'de baja o fin de contrato si la hay (ese número es el que va '
                'al finiquito) y, si alguien tiene un derecho distinto del '
                'general, sus días propios.')
    ws['A3'].font = Font(size=9, italic=True)
    ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[3].height = 28

    # RD-27 · el prorrateo sólo funcionaba por un extremo: se prorrateaba el
    # ALTA y no la BAJA. Un temporal o un fijo-discontinuo de enero a junio
    # conservaba los 30 días enteros de derecho cuando le corresponden 15 — y
    # 15 es justo el número que se liquida en el finiquito. El 07 ya tiene
    # «Fecha Fin Contrato» (columna L) y su DV contempla Temporal, Prácticas,
    # Formación y Fijo-discontinuo; aquí no había dónde escribirla.
    _cabecera(ws, 4, ['Empleado', 'Fecha de alta',
                      'Fecha de baja / fin de contrato', 'Días propios',
                      'Derecho del año', 'Disfrutados', 'Pendientes',
                      'Le quedan', 'Aviso'])
    for i in range(motor.CAPACIDAD):
        f, cal = 5 + i, EMP0 + i
        _formula(ws, 'A{}'.format(f),
                 "=IF('Calendario Anual'!$A{c}=\"\",\"\","
                 "'Calendario Anual'!$A{c})".format(c=cal))
        _verde(ws, 'B{}'.format(f), fmt=motor.FMT_FECHA)
        _verde(ws, 'C{}'.format(f), fmt=motor.FMT_FECHA)
        _verde(ws, 'D{}'.format(f), fmt=motor.FMT_ENT)
        # Derecho: el general (B2) o el propio (D), prorrateado por los días
        # de PERMANENCIA dentro del ejercicio.
        #
        #   · inicio = MAX(fecha de alta, 1 de enero)      → $B{f} o $E$2
        #   · fin    = MIN(fecha de baja, cierre del año)  → $C{f} o $G$2
        #
        # RT-08 · y con suelo: la versión anterior cubría por ARRIBA con MIN
        # y no por abajo, así que una fecha de alta posterior al cierre del
        # año —trivial de teclear— producía un DERECHO NEGATIVO que se
        # arrastraba a «Le quedan» y disparaba «⛔ saldo negativo», que el
        # cliente lee como un exceso de vacaciones. El divisor tampoco puede
        # ser un 365 cableado: sale del propio ejercicio, y así no falla en
        # bisiesto.
        derecho = 'IF($D{f}="",$B$2,$D{f})'.format(f=f)
        ini = 'IF($B{f}="",$E$2,MAX($B{f},$E$2))'.format(f=f)
        fin = 'IF($C{f}="",$G$2,MIN($C{f},$G$2))'.format(f=f)
        _formula(ws, 'E{}'.format(f),
                 '=IFERROR(IF($A{f}="","",'
                 'MAX(0,MIN({d},ROUND({d}*(({fin})-({ini})+1)/'
                 '($G$2-$E$2+1),1)))),"")'
                 .format(f=f, d=derecho, ini=ini, fin=fin), motor.FMT_DEC1)
        for col, estado in (('F', 'Aprobado'), ('G', 'Pendiente')):
            _formula(ws, '{}{}'.format(col, f),
                     '=IF($A{f}="","",{s})'
                     .format(f=f,
                             s=_sumifs_solicitudes('$A{}'.format(f), estado)),
                     motor.FMT_DEC1)
        _formula(ws, 'H{}'.format(f),
                 '=IF($E{f}="","",ROUND($E{f}-$F{f}-$G{f},1))'.format(f=f),
                 motor.FMT_DEC1)
        # RT-08 · el aviso explícito de la fecha imposible, que es lo que
        # convierte un número raro en algo accionable.
        _formula(ws, 'I{}'.format(f),
                 '=IF($A{f}="","",'
                 'IF(AND($B{f}<>"",$B{f}>$G$2),'
                 '"⚠ la fecha de alta es posterior al cierre del año",'
                 'IF(AND($C{f}<>"",$B{f}<>"",$C{f}<$B{f}),'
                 '"⚠ la fecha de baja es anterior a la de alta",'
                 'IF($H{f}<0,"⛔ saldo negativo",'
                 'IF($H{f}=0,"⚠ sin días disponibles","")))))'.format(f=f))
    _cf(ws, 'I5:I{}'.format(4 + motor.CAPACIDAD), motor.VOC_ALERTA)

    fin = 4 + motor.CAPACIDAD
    ws['A{}'.format(fin + 2)] = (
        '▸ «Le quedan» descuenta las solicitudes APROBADAS y también las '
        'PENDIENTES: es el saldo que puedes comprometer hoy sin aprobar dos '
        'veces los mismos días. Y sólo cuenta las de ESTE ejercicio (las '
        'fechas de arriba), así que el mismo libro sirve varios años sin '
        'falsear el saldo.')
    ws['A{}'.format(fin + 3)] = (
        '▸ El prorrateo es lineal sobre los días que la persona está de alta '
        'DENTRO del ejercicio: entra la fecha de alta y también la de baja o '
        'fin de contrato. Un temporal de enero a junio genera 15 días, no 30 '
        '— y ése es el número que va al finiquito. Tu convenio puede '
        'calcularlo por meses completos: en ese caso escribe los días que '
        'correspondan en «Días propios».')
    for r in (fin + 2, fin + 3):
        ws.cell(row=r, column=1).font = Font(size=9, italic=True)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True,
                                                       vertical='top')
        ws.row_dimensions[r].height = 30
    ws['A{}'.format(fin + 5)] = motor.PIE
    ws['A{}'.format(fin + 5)].font = Font(size=8, color='888888')
    _anchos(ws, {'A': 24, 'B': 16, 'C': 22, 'D': 14, 'E': 16, 'F': 14,
                 'G': 14, 'H': 14, 'I': 34})
    _merges(ws, ['A1:I1', 'A3:I3', 'A{0}:I{0}'.format(fin + 2),
                 'A{0}:I{0}'.format(fin + 3),
                 'A{0}:I{0}'.format(fin + 5)])
    ws.freeze_panes = 'A5'
    cambios.append('05:Saldo Vacaciones!{}: días de convenio en UNA celda '
                   '(antes en 30 fórmulas), prorrateo por fecha de alta y '
                   'saldo que descuenta la solicitud en curso '
                   '(DOM-19/TEC-18/COM-24)'.format(coord))
    cambios.append('05:Saldo Vacaciones!C4:C34: columna nueva «Fecha de baja '
                   '/ fin de contrato». El prorrateo sólo funcionaba por un '
                   'extremo: un temporal de enero a junio conservaba los 30 '
                   'días enteros cuando le corresponden 15, que es el número '
                   'que se liquida en el finiquito — RD-27')
    cambios.append('05:Saldo Vacaciones!E5:E34/I5:I34: el derecho lleva '
                   'MAX(0, MIN(...)) —una fecha de alta posterior al cierre '
                   'del año daba derecho NEGATIVO y disparaba «⛔ saldo '
                   'negativo», que se lee como exceso de vacaciones— y el '
                   'divisor sale del propio ejercicio en vez de un 365 '
                   'cableado, que falla en bisiesto. La columna «Aviso» '
                   'nombra las dos fechas imposibles — RT-08')
    cambios.append('05:Saldo Vacaciones!E2/G2 y F5:G34: los SUMIFS filtran '
                   'por AÑO contra la fecha de inicio de la solicitud. Sin '
                   'ese filtro, reutilizar el libro al año siguiente hacía '
                   'que las vacaciones de dos ejercicios descontaran del '
                   'mismo derecho de 30 días y toda la plantilla apareciera '
                   'en negativo — RD-10')


def _n05_solicitudes(wb, cambios):
    """§4 · DOM-20 — fechas invertidas, saldo por INDEX/MATCH y verificación
    de que el nombre existe en el calendario.

    `D5='=IF(AND(B5<>"",C5<>""),C5-B5+1,"")'` daba **-4 días** con las fechas
    al revés, y ese número entraba tal cual en el saldo.
    """
    ws = wb['Solicitudes']
    ult = SOL1
    saldo_ult = 4 + motor.CAPACIDAD
    # Las combinaciones se deshacen ANTES de escribir: el pie de la v1.1 va
    # combinado (A36:H36) y ahora ahí hay filas de solicitud — escribir sobre
    # una `MergedCell` revienta con «value is read-only».
    _merges(ws, [])
    _vaciar(ws, 5, max(ult + 4, ws.max_row))
    _cabecera(ws, 4, ['Empleado', 'Fecha Inicio', 'Fecha Fin',
                      'Días solicitados', 'Ya disfrutados', 'Le quedan',
                      'Estado', 'Aviso'])
    for f in range(SOL0, ult + 1):
        _formula(ws, 'D{}'.format(f),
                 '=IF(OR($B{f}="",$C{f}=""),"",IF($C{f}<$B{f},'
                 '"⚠ fechas invertidas",$C{f}-$B{f}+1))'.format(f=f))
        # El saldo se lee del bloque de 30 EMPLEADOS de «Saldo Vacaciones»
        # (que sí está indexado por persona), no de las 150 peticiones.
        for col, origen in (('E', 'F'), ('F', 'H')):
            _formula(ws, '{}{}'.format(col, f),
                     "=IFERROR(IF($A{f}=\"\",\"\",INDEX("
                     "'Saldo Vacaciones'!${o}$5:${o}${u},MATCH($A{f},"
                     "'Saldo Vacaciones'!$A$5:$A${u},0))),\"\")"
                     .format(f=f, o=origen, u=saldo_ult), motor.FMT_DEC1)
        _formula(ws, 'H{}'.format(f),
                 '=IF($A{f}="","",IF(COUNTIF(\'Calendario Anual\'!$A${a}:$A${b}'
                 ',$A{f}&"")=0,"⚠ ese nombre no está en el calendario",""))'
                 .format(f=f, a=EMP0, b=EMP1))
    _cf(ws, 'D{}:D{}'.format(SOL0, ult), motor.VOC_ALERTA)
    _cf(ws, 'H{}:H{}'.format(SOL0, ult), motor.VOC_ALERTA)
    _anchos(ws, {'A': 24, 'B': 14, 'C': 14, 'D': 16, 'E': 14, 'F': 12,
                 'G': 14, 'H': 30})
    nota = ult + 2
    ws['A{}'.format(nota)] = (
        '▸ Esta hoja se indexa por PETICIÓN, no por persona: caben {} '
        'solicitudes, unas cinco por empleado. Nadie coge los 30 días del '
        'tirón — en hostelería se parten en verano, Navidad y algún puente—, '
        'y el saldo sólo es cierto si TODOS los periodos están aquí.'
        .format(SOL_FILAS))
    ws['A{}'.format(nota)].font = Font(size=9, italic=True)
    ws['A{}'.format(nota)].alignment = Alignment(wrap_text=True,
                                                 vertical='top')
    ws.row_dimensions[nota].height = 30
    ws['A{}'.format(nota + 2)] = motor.PIE
    ws['A{}'.format(nota + 2)].font = Font(size=8, color='888888')
    _merges(ws, ['A1:H1', 'A2:H2', 'A{0}:H{0}'.format(nota),
                 'A{0}:H{0}'.format(nota + 2)])
    cambios.append('05:Solicitudes!D{}:H{}: aviso de fechas invertidas, saldo '
                   'por INDEX/MATCH sobre «Saldo Vacaciones» y verificación '
                   'del nombre contra el calendario (DOM-20)'
                   .format(SOL0, ult))
    cambios.append('05:Solicitudes: {} → {} filas. La hoja NO está indexada '
                   'por empleado sino por PETICIÓN, y con 30 filas para 30 '
                   'personas era UNA solicitud por persona y año: partir las '
                   'vacaciones en dos o tres periodos la agotaba a los 10-15 '
                   'empleados y a partir de ahí el saldo dejaba de ser cierto '
                   'en silencio, porque los SUMIFS y los INDEX/MATCH tenían '
                   'el rango cerrado en la fila 34 — RD-09/RT-03'
                   .format(motor.CAPACIDAD, SOL_FILAS))


def _n05_cobertura(wb, cambios):
    """§4 · DOM-20/COM-20 — `Cobertura` deja de ser una tabla vacía.

    La hoja tenía dos cabeceras y CERO datos, mientras la landing prometía
    «cobertura mínima» y «periodos de máxima demanda».
    """
    ws = wb['Cobertura']
    motor._limpiar_dv(ws)
    _merges(ws, [])            # el pie de la v1.1 va combinado en A29:E29
    fin_sust = 11 + motor.CAPACIDAD                      # 41
    _vaciar(ws, 28)

    _rotulo(ws, 3, 'Máximo de ausencias simultáneas por semana:')
    _verde(ws, 'B3', 4, motor.FMT_ENT)
    # RT-11 · el umbral de la temporada alta era un `>0` fijo dentro de la
    # fórmula: UNA sola ausencia en cualquiera de las 11 semanas precargadas
    # como «Alta» pintaba el bloqueante ⛔, así que el usuario aprendía a
    # ignorarlo. Ahora es un parámetro, como manda §1.4.
    _rotulo(ws, 3, 'Máximo en TEMPORADA ALTA:', col=3)
    _verde(ws, 'D3', 1, motor.FMT_ENT)
    _seccion(ws, 4, 'PERSONAL MÍNIMO POR TURNO', 6)
    _cabecera(ws, 5, ['Turno', 'Mínimo Personal', 'Cocina', 'Sala', 'Barra'])
    # RD-14 · los mínimos precargados contradecían al resto del kit: la tarde
    # pedía 6 personas mientras 03!'Previsión por Servicio'!E18 y
    # BONUS-02!E26 calculaban 5 para el mismo caso por defecto, y la NOCHE
    # (23:00-07:00 según 01!'Turnos') pedía 5, que no corresponde a un casual
    # de 80 cubiertos. Se alinean con lo que calculan los otros dos ficheros:
    # el turno de servicio son las 5 personas del dimensionado, la mañana es
    # montaje y comidas y la noche es cierre.
    for i, fila in enumerate([('M · Mañana', 2, 2, 0), ('T · Tarde', 2, 2, 1),
                              ('N · Noche (cierre)', 1, 1, 0)]):
        f = 6 + i
        ws['A{}'.format(f)] = fila[0]
        for j, col in enumerate(('C', 'D', 'E')):
            _verde(ws, '{}{}'.format(col, f), fila[1 + j], motor.FMT_ENT)
        _formula(ws, 'B{}'.format(f),
                 '=IF(COUNT($C{f}:$E{f})=0,"",$C{f}+$D{f}+$E{f})'.format(f=f),
                 motor.FMT_ENT)
    ws['A9'] = ('▸ Por encima del máximo de B3 la fila «Cobertura» del '
                'calendario salta en ámbar; en temporada alta el umbral es el '
                'de D3. Los mínimos por turno son los que calculan el fichero '
                '03 y el BONUS-02 para el caso por defecto (5 personas por '
                'servicio): son un punto de partida, ajústalos a tu carta. Y '
                'no son decorativos: el último indicador de abajo cuenta las '
                'semanas en que la plantilla disponible no llega al turno más '
                'exigente.')
    ws['A9'].font = Font(size=9, italic=True)
    ws['A9'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[9].height = 42

    _seccion(ws, 10, 'COBERTURA: QUIÉN SUSTITUYE A QUIÉN', 6)
    _cabecera(ws, 11, ['Empleado Ausente', 'Puesto', 'Sustituto 1',
                       'Sustituto 2', 'Notas', 'Días pendientes', 'Aviso'])
    for f in range(12, fin_sust + 1):
        _formula(ws, 'F{}'.format(f),
                 "=IFERROR(IF($A{f}=\"\",\"\",INDEX('Saldo Vacaciones'!"
                 "$H$5:$H${u},MATCH($A{f},'Saldo Vacaciones'!$A$5:$A${u},0))"
                 "),\"\")".format(f=f, u=4 + motor.CAPACIDAD),
                 motor.FMT_DEC1)
        # RT-18 · «Empleado Ausente» era texto libre y su «Días pendientes»
        # resolvía por INDEX/MATCH envuelto en IFERROR: un nombre mal escrito
        # devolvía celda VACÍA, indistinguible de alguien sin días
        # pendientes. La hoja hermana 'Solicitudes' sí traía el aviso para el
        # mismo riesgo y la columna B de ESTA hoja sí trae DV: la
        # incoherencia estaba dentro de la misma pantalla.
        _formula(ws, 'G{}'.format(f),
                 '=IF($A{f}="","",IF(COUNTIF(\'Calendario Anual\'!$A${a}:$A${b}'
                 ',$A{f}&"")=0,"⚠ ese nombre no está en el calendario",""))'
                 .format(f=f, a=EMP0, b=EMP1))
    _cf(ws, 'G12:G{}'.format(fin_sust), motor.VOC_ALERTA)
    _dv(ws, 'B12:B{}'.format(fin_sust), PUESTOS, 'Puesto no válido',
        'El puesto decide quién puede sustituir a quién: un ayudante de sala '
        'no cubre una jefatura de cocina. Si te falta alguno, desprotege la '
        'hoja y edita la lista.', 'Elige el puesto de la persona ausente.')

    _seccion(ws, 43, 'INDICADORES DE COBERTURA', 7)
    indicadores = [
        (44, 'Plantilla total en el calendario:',
         "=COUNTIF('Calendario Anual'!$A${a}:$A${b},\"?*\")"),
        (45, 'Pico de ausencias simultáneas (peor semana):',
         "=MAX('Calendario Anual'!$B${m}:${c}${m})"),
        (46, 'Semanas por encima de tu máximo:',
         "=COUNTIF('Calendario Anual'!$B${m}:${c}${m},\">\"&$B$3)"),
        (47, 'Semanas de temporada alta con alguien ausente:',
         "=SUMPRODUCT(('Calendario Anual'!$B${t}:${c}${t}=\"Alta\")*"
         "('Calendario Anual'!$B${m}:${c}${m}>0))"),
        # RD-14/RT-14 · el bloque «PERSONAL MÍNIMO POR TURNO» era decorativo:
        # ninguna fórmula del kit lo leía —ni los indicadores de esta hoja ni
        # la alerta del calendario, que sólo miraba B3—, así que el cliente
        # que ajustara esos mínimos no veía cambiar nada en ningún sitio. Y
        # encima contradecía al 03 y al BONUS-02 con cifras inventadas. Este
        # indicador lo CONECTA: cuenta las semanas en que la plantilla que
        # queda disponible no llega a cubrir el turno más exigente.
        (48, 'Semanas en que NO se cubre el turno más exigente '
             '(plantilla − ausencias < mínimo):',
         "=SUMPRODUCT((($B$44-'Calendario Anual'!$B${m}:${c}${m})"
         "<MAX($B$6:$B$8))*1)"),
    ]
    for fila, rotulo, formula in indicadores:
        _rotulo(ws, fila, rotulo)
        _formula(ws, 'B{}'.format(fila),
                 formula.format(a=EMP0, b=EMP1, m=F_MAX, t=F_TEMP, c=COL_SN),
                 motor.FMT_ENT)

    ws['A50'] = ('▸ Los indicadores leen la fila «Ausencias en la semana» del '
                 'Calendario Anual, que ahora es la MAYOR de dos cuentas: lo '
                 'que hayas pintado a mano en la rejilla (V, B y PE) y lo que '
                 'sale solo de las solicitudes APROBADAS. Trabajes con la '
                 'rejilla o con las solicitudes, la cobertura se entera.')
    ws['A50'].font = Font(size=9, italic=True)
    ws['A50'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[50].height = 34
    ws['A52'] = motor.PIE
    ws['A52'].font = Font(size=8, color='888888')
    _anchos(ws, {'A': 46, 'B': 18, 'C': 26, 'D': 12, 'E': 26, 'F': 16,
                 'G': 30})
    _merges(ws, ['A1:G1', 'A2:G2', 'A9:G9', 'A50:G50', 'A52:G52'])
    cambios.append('05:Cobertura!B3/B44:B47: máximo de ausencias en celda, '
                   'personal mínimo por turno precargado e indicadores reales '
                   'de pico y temporada alta (COM-20)')
    cambios.append('05:Cobertura!D3: umbral propio de la TEMPORADA ALTA en '
                   'celda verde (1 por defecto). Antes la rama de temporada '
                   'alta se evaluaba primero y con «>0», así que una sola '
                   'ausencia en cualquiera de las 11 semanas precargadas '
                   'pintaba el bloqueante ⛔ y el «⚠ EXCESO» —la única alerta '
                   'que compara contra el máximo del cliente— no aparecía '
                   'nunca — RT-11')
    cambios.append('05:Cobertura!B48: indicador nuevo que CONECTA el bloque '
                   '«Personal mínimo por turno», hasta ahora decorativo: '
                   'cuenta las semanas en que la plantilla disponible no '
                   'llega al turno más exigente. Y los mínimos precargados '
                   'dejan de contradecir al 03 y al BONUS-02 — RD-14/RT-14')
    cambios.append('05:Cobertura!G12:G41: aviso «⚠ ese nombre no está en el '
                   'calendario», el mismo que ya tenía «Solicitudes». Un '
                   'nombre mal escrito dejaba «Días pendientes» en blanco, '
                   'indistinguible de alguien sin días pendientes — RT-18')


def _n05_instrucciones(wb, cambios):
    ws = wb['Instrucciones']
    _instrucciones(ws, [
        '05 · Planificación de Vacaciones Anual',
        None,
        'Cómo usar esta plantilla:',
        "▸ Escribe los nombres una sola vez, en la columna A del 'Calendario "
        "Anual'. El resto de hojas los leen de ahí.",
        '▸ La rejilla es de SEMANAS, no de meses: la fila 4 numera las 53 '
        'semanas del año y la fila 5 enseña el lunes de cada una. Cambia el '
        'lunes de la semana 1 (celda verde B5) y las 52 restantes se '
        'recalculan solas.',
        '▸ ' + motor.LEYENDA_AUSENCIA + '.',
        "▸ Registra cada petición en 'Solicitudes' con sus fechas de inicio y "
        'fin. Los días se cuentan SOLOS y avisan si las fechas están al revés.',
        "▸ El saldo real vive en 'Saldo Vacaciones': derecho del año, "
        'disfrutados, pendientes y lo que le queda a cada persona. Descuenta '
        'también las solicitudes pendientes, para que no apruebes dos veces '
        'los mismos días.',
        "▸ En 'Cobertura' fijas el personal mínimo por turno y el máximo de "
        'ausencias que aguantas a la vez. Con eso, la fila «Cobertura» del '
        'calendario avisa semana a semana.',
        None,
        'Por qué el calendario ya no cuenta meses:',
        '▸ En la versión 1.1 había una casilla por MES, así que quien '
        'disfrutaba agosto entero figuraba con 1 día usado y 29 restantes. '
        'Ahora la columna «Días Usados» suma los DÍAS de las solicitudes '
        'aprobadas de esa persona: la rejilla sirve para VER el año de un '
        'vistazo y las solicitudes son la fuente del cómputo.',
        None,
        'Legislación española:',
        '▸ 30 días naturales de vacaciones al año (art. 38 ET). El número '
        "está en UNA celda verde ('Saldo Vacaciones', B2): si tu convenio "
        'habla de 22 días laborables, escribe 22 y cuenta sólo días de '
        'trabajo en las solicitudes.',
        '▸ El calendario de vacaciones se fija de común acuerdo y se conoce '
        'con al menos dos meses de antelación a su disfrute (art. 38.3 ET).',
        '▸ Las vacaciones no se sustituyen por dinero salvo al liquidar el '
        'contrato.',
        None,
        'Aviso: tu convenio de hostelería manda sobre los mínimos del '
        'Estatuto y puede mejorar cualquiera de estos números.',
    ])
    cambios.append('05:Instrucciones!B2:B24: reescritas — rejilla por '
                   'semanas, hoja «Saldo Vacaciones» y el 30 como celda '
                   'editable, no como constante de 30 fórmulas')


# ==========================================================================
# Contrato con main.py
# ==========================================================================
def pre(wb, fname, cambios):
    """Todo el trabajo del grupo, ANTES de `motor.aplicar` (ver la cabecera)."""
    if fname.startswith('03-'):
        _n03_nominas(wb, cambios)
        _n03_ratio(wb, cambios)
        _n03_prevision(wb, cambios)
        _n03_instrucciones(wb, cambios)
    elif fname.startswith('04-'):
        _n04(wb, cambios)
        _n04_instrucciones(wb, cambios)
    elif fname.startswith('05-'):
        _n05_calendario(wb, cambios)
        _n05_saldo(wb, cambios)
        _n05_solicitudes(wb, cambios)
        _n05_cobertura(wb, cambios)
        _n05_instrucciones(wb, cambios)
    return cambios


# ==========================================================================
# Demostraciones con pycel (SPEC §5) — se cambian ENTRADAS y se comprueba la
# DIRECCIÓN del resultado, que es lo único que demuestra que la cadena existe.
# ==========================================================================
import contextlib                                          # noqa: E402
import os                                                  # noqa: E402

EPOCA = datetime.datetime(1899, 12, 30)


def _serie(y, m, d):
    """Fecha → número de serie de Excel. pycel opera con el serial, no con el
    `datetime`: pasarle un objeto fecha en `set_value` deja la resta de
    `C5-B5+1` en `#VALUE!`."""
    return (datetime.datetime(y, m, d) - EPOCA).days


def _xl(path):
    from pycel import ExcelCompiler
    return ExcelCompiler(filename=path)


def _ev(xl, ref):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            return xl.evaluate(ref)
        except Exception as e:                              # noqa: BLE001
            return 'ERR:{}'.format(type(e).__name__)


def _sv(xl, ref, valor):
    with open(os.devnull, 'w') as dn, contextlib.redirect_stderr(dn):
        try:
            xl.evaluate(ref)
            xl.set_value(ref, valor)
            return True
        except Exception:                                   # noqa: BLE001
            return False


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _calentar(xl, refs):
    """⚠ pycel sólo propaga un `set_value` a las celdas que YA están en su
    grafo de dependencias. Medido el 2026-08-24 con la fila de cobertura del
    05: escribir «V» en `Calendario Anual!AB6` sin haber evaluado antes nada
    de esa columna dejaba `AB37` valiendo 0 y el indicador
    `Cobertura!B47` (SUMPRODUCT sobre las 53 semanas) en 0 — la demostración
    «fallaba» por un artefacto del evaluador, no por la hoja. Evaluando primero
    las SALIDAS, la cadena entera entra en el grafo y la invalidación funciona.
    """
    return dict((r, _ev(xl, r)) for r in refs)


def _d03_coste(carpeta):
    """§3 · la cotización, las pagas y el coste/hora salen de CELDAS."""
    p = os.path.join(carpeta, '03-coste-laboral-mensual.xlsx')
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    ref = lambda c: "'Nóminas'!{}".format(c)                # noqa: E731
    vacia = dict((c, _ev(xl, ref(c + '5'))) for c in 'EFGI')
    _sv(xl, ref('C5'), 1500)
    _sv(xl, ref('D5'), 14)
    _sv(xl, ref('H5'), 40)
    base = dict((c, _ev(xl, ref(c + '5'))) for c in 'EFGI')
    _sv(xl, ref('F2'), 52)
    base['I_52'] = _ev(xl, ref('I5'))
    _sv(xl, ref('F2'), SEMANAS_EFECTIVAS)
    _sv(xl, ref('D5'), 12)
    doce = _ev(xl, ref('E5'))
    _sv(xl, ref('D5'), 14)
    _sv(xl, ref('C2'), 0.40)
    ss40 = _ev(xl, ref('F5'))
    ok = (all(v in ('', None) for v in vacia.values())
          and _num(base['E']) and abs(base['E'] - 1750.0) < 0.01
          and _num(base['F']) and abs(base['F'] - 577.5) < 0.01
          and _num(base['G']) and abs(base['G'] - 2327.5) < 0.01
          # RD-20 · el divisor son las semanas EFECTIVAS (46,5), no 52:
          # 2.327,50 / (40 × 46,5 / 12) = 15,02 €/h. Con las 52 del
          # calendario salían 13,43 €/h — un 11 % corto, y ese número es el
          # que las Instrucciones mandan llevar a la tarifa del 02.
          and _num(base['I']) and abs(base['I'] - 15.02) < 0.02
          and _num(base['I_52']) and abs(base['I_52'] - 13.43) < 0.02
          and _num(doce) and abs(doce - 1500.0) < 0.01
          and _num(ss40) and ss40 > base['F'])
    return {
        'ref': '03-coste-laboral-mensual.xlsx:Nóminas:C2/E5/F5/G5/I5',
        'fila_vacia': dict((k, str(v)) for k, v in vacia.items()),
        'bruto_1500_x14_pagas': {'prorrateado_E5': base['E'],
                                 'cotizacion_F5': base['F'],
                                 'coste_total_G5': base['G'],
                                 'coste_hora_I5_con_46_5_semanas': base['I'],
                                 'coste_hora_I5_con_52_semanas':
                                     base['I_52']},
        'mismo_bruto_en_12_pagas_E5': doce,
        'cotizacion_al_40_por_ciento_F5': ss40,
        'ok': ok,
        'nota': 'la v1.1 hacía `=IF(C5<>"",C5*0.30,"")` sin pagas: 1.500 € en '
                '14 pagas costaban 1.950 €/mes en vez de 2.327,50 € — un 19 % '
                'de error a favor del optimismo (DOM-08/DOM-31/TEC-14). Y el '
                'coste/hora dividía entre las 52 semanas del calendario, '
                'contando como trabajadas las de vacaciones y los festivos: '
                '13,43 €/h en vez de 15,02 (RD-20)'}


def _d03_semaforo(carpeta):
    """§3 + RD-01/RT-01/RD-21/RC-11 — el semáforo del coste laboral.

    Cuatro cosas se miden aquí, todas cambiando entradas sobre MI copia:

      1. El MISMO ratio del 40 % es rojo en un fast casual y ámbar en un fine
         dining, porque los umbrales salen de la tabla por VLOOKUP (DOM-16).
      2. **RD-01/RT-01**: teclear las VENTAS sin una sola nómina ya no produce
         «🟢 EXCELENTE». `B5` es un `SUM` sobre una hoja vacía, así que vale 0
         y no `""`: la guarda tenía que mirar el COSTE, no sólo las ventas.
      3. **RD-21**: horas extra, ETT y otros costes de personal entran en el
         coste y mueven el ratio.
      4. **RC-11**: «Hotel (restaurante)» existe en el 03 y devuelve los
         MISMOS 42/44 % que el BONUS-02.
    """
    p = os.path.join(carpeta, '03-coste-laboral-mensual.xlsx')
    if not os.path.isfile(p):
        return None
    hoja = "'Ratio Coste Laboral'!"

    def escena(sets, lecturas=('B11', 'B13', 'D11', 'D12', 'B9')):
        # ⚠ las SALIDAS se evalúan ANTES de escribir las entradas: pycel sólo
        # propaga un `set_value` a las celdas que ya están en su grafo de
        # dependencias (ver `_calentar`). Compilador limpio por escena para no
        # arrastrar estado entre casos.
        xl = _xl(p)
        _calentar(xl, [hoja + c for c in lecturas])
        for ref, val in sets:
            _calentar(xl, [hoja + ref])
            _sv(xl, hoja + ref, val)
        return dict((c, _ev(xl, hoja + c)) for c in lecturas)

    fresco = escena([])
    # RD-01 · el caso exacto del hallazgo: ventas sí, nóminas no
    solo_ventas = escena([('B4', 50000)])
    # el coste entra por las nóminas (B5 es fórmula: se fuerza el valor)
    base = escena([('B4', 50000), ('B5', 20000)])
    fine = escena([('B4', 50000), ('B5', 20000),
                   ('B3', 'Fine Dining / Alta Cocina')])
    fast = escena([('B4', 50000), ('B5', 20000),
                   ('B3', 'Fast Casual / Comida Rápida')])
    hotel = escena([('B4', 50000), ('B5', 20000),
                    ('B3', 'Hotel (restaurante)')])
    # RD-21 · las tres casillas nuevas suman al coste
    con_extras = escena([('B4', 50000), ('B5', 20000), ('B6', 1200),
                         ('B7', 800), ('B8', 500)])
    # RD-01 · suelo de plausibilidad: 5 % de ratio con objetivo del 33 %
    implausible = escena([('B4', 50000), ('B5', 2500)])

    ok = (fresco['B11'] in ('', None)
          and 'VENTAS' in str(fresco['B13'])
          and solo_ventas['B11'] in ('', None)
          and 'nóminas' in str(solo_ventas['B13'])
          and abs((base['B11'] or 0) - 0.4) < 0.0001
          and 'ACCIÓN CORRECTIVA' in str(base['B13'])
          and abs((fine['D11'] or 0) - 0.40) < 0.0001
          and 'VIGILAR' in str(fine['B13'])
          and abs((fast['D12'] or 0) - 0.30) < 0.0001
          and 'ACCIÓN CORRECTIVA' in str(fast['B13'])
          and abs((hotel['D11'] or 0) - 0.42) < 0.0001
          and abs((hotel['D12'] or 0) - 0.44) < 0.0001
          and 'EXCELENTE' in str(hotel['B13'])
          and abs((con_extras['B9'] or 0) - 22500) < 0.01
          and abs((con_extras['B11'] or 0) - 0.45) < 0.0001
          and 'implausible' in str(implausible['B13']))
    return {'ref': '03-coste-laboral-mensual.xlsx:Ratio Coste Laboral:'
                   'B9/B11/B13',
            'hoja_recien_descargada': {'ratio': str(fresco['B11']),
                                       'veredicto': str(fresco['B13'])},
            'ventas 50.000 SIN una sola nómina (RD-01/RT-01)':
                {'coste_B9': str(solo_ventas['B9']),
                 'ratio': str(solo_ventas['B11']),
                 'veredicto': str(solo_ventas['B13'])},
            'ratio 40 % sin tipo elegido': {'ratio': base['B11'],
                                            'veredicto': base['B13']},
            'ratio 40 % en fine dining': {'objetivo': fine['D11'],
                                          'maximo': fine['D12'],
                                          'veredicto': fine['B13']},
            'ratio 40 % en fast casual': {'objetivo': fast['D11'],
                                          'maximo': fast['D12'],
                                          'veredicto': fast['B13']},
            'ratio 40 % en HOTEL, que antes no existía en el 03 (RC-11)':
                {'objetivo': hotel['D11'], 'maximo': hotel['D12'],
                 'veredicto': hotel['B13']},
            'con horas extra 1.200 + ETT 800 + otros 500 (RD-21)':
                {'coste_total_B9': con_extras['B9'],
                 'ratio': con_extras['B11'],
                 'veredicto': con_extras['B13']},
            'coste 2.500 sobre ventas de 50.000 = 5 % (RD-01)':
                {'ratio': implausible['B11'],
                 'veredicto': implausible['B13']},
            'ok': ok,
            'nota': 'la v1.1 daba «🟢 EXCELENTE (<30%)» con la hoja vacía y '
                    'suspendía a cualquiera por encima del 35 %, incluida la '
                    'alta cocina que su propia tabla declara correcta hasta '
                    'el 42 %. Y la v2.0 seguía dando EXCELENTE en cuanto se '
                    'tecleaban las ventas sin nóminas, porque la guarda sólo '
                    'miraba B4 (DOM-16/COM-11/COM-17/RD-01/RT-01/RD-21/RC-11)'}


def _d03_fte(carpeta):
    """§3 · la cadena de dimensionamiento por SERVICIO: 7 FTE por defecto, y
    reacciona a los cubiertos, a los servicios y a un ratio a cero."""
    p = os.path.join(carpeta, '03-coste-laboral-mensual.xlsx')
    if not os.path.isfile(p):
        return None
    xl = _xl(p)
    h = "'Previsión por Servicio'!"
    base = dict(cubiertos_por_servicio_B6=_ev(xl, h + 'B6'),
                cocina_B18=_ev(xl, h + 'B18'), sala_C18=_ev(xl, h + 'C18'),
                barra_D18=_ev(xl, h + 'D18'),
                por_servicio_E18=_ev(xl, h + 'E18'),
                presencias_B19=_ev(xl, h + 'B19'),
                horas_semana_B20=_ev(xl, h + 'B20'),
                fte_B21=_ev(xl, h + 'B21'), coste_B22=_ev(xl, h + 'B22'))
    _sv(xl, h + 'B4', 160)
    doble = {'fte_B21': _ev(xl, h + 'B21'), 'coste_B22': _ev(xl, h + 'B22')}
    _sv(xl, h + 'B4', 80)
    _sv(xl, h + 'B11', 7)
    siete_dias = _ev(xl, h + 'B21')
    _sv(xl, h + 'B11', 6)
    _sv(xl, h + 'B9', 0)
    sin_barra = {'barra_D18': _ev(xl, h + 'D18'),
                 'por_servicio_E18': _ev(xl, h + 'E18')}
    ok = (base['cubiertos_por_servicio_B6'] == 40
          and base['por_servicio_E18'] == 5
          and base['presencias_B19'] == 10
          and base['horas_semana_B20'] == 240
          and base['fte_B21'] == 7
          and _num(base['coste_B22'])
          and abs(base['coste_B22'] - 16292.50) < 0.01
          and _num(doble['fte_B21']) and doble['fte_B21'] > 7
          and _num(siete_dias) and siete_dias > 7
          and sin_barra['barra_D18'] == 0
          and sin_barra['por_servicio_E18'] == 4)
    return {'ref': '03-coste-laboral-mensual.xlsx:Previsión por Servicio:B21',
            'por_defecto_80_cubiertos_2_servicios_casual': base,
            'al_doblar_los_cubiertos_a_160': doble,
            'abriendo_7_dias_en_vez_de_6_fte': siete_dias,
            'con_el_ratio_de_barra_a_cero': sin_barra,
            'ok': ok,
            'nota': 'la v1.1 dividía los cubiertos del DÍA entre el ratio y '
                    'daba 15 personas por turno para 80 cubiertos; y sus '
                    '«días de apertura» no entraban en ninguna fórmula '
                    '(DOM-09/DOM-18/TEC-15/COM-09)'}


def _d04_progreso(carpeta):
    """§4 · 0 % recién descargado, y la marca «—» sale del denominador."""
    p = os.path.join(carpeta, '04-onboarding-nuevo-empleado.xlsx')
    if not os.path.isfile(p):
        return None
    import openpyxl
    ws = openpyxl.load_workbook(p)['Checklist Onboarding']
    tramos = [(r0, r1) for _t, _h, r0, r1 in motor.secciones_04(ws)]
    fila = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().upper() == 'RESUMEN':
            fila = r
            break
    if fila is None:
        return {'ok': False, 'nota': 'no se encuentra el bloque RESUMEN'}
    h = "'Checklist Onboarding'!"
    c1, c2, c3 = 'C{}'.format(fila + 1), 'C{}'.format(fila + 2), \
                 'C{}'.format(fila + 3)
    tot = 'E{}'.format(fila + 1)
    xl = _xl(p)
    fresco = {'completadas': _ev(xl, h + c1), 'aplicables': _ev(xl, h + c2),
              'total': _ev(xl, h + tot), 'progreso': _ev(xl, h + c3)}
    _sv(xl, h + 'F{}'.format(tramos[0][0]), '✓')
    _sv(xl, h + 'F{}'.format(tramos[1][0]), '✓')
    dos = {'completadas': _ev(xl, h + c1), 'progreso': _ev(xl, h + c3)}
    _sv(xl, h + 'F{}'.format(tramos[2][0]), '—')
    noaplica = {'aplicables': _ev(xl, h + c2), 'progreso': _ev(xl, h + c3)}
    total_tareas = sum(b - a + 1 for a, b in tramos)
    ok = (fresco['completadas'] == 0 and fresco['progreso'] == 0
          and fresco['total'] == total_tareas == 50
          and fresco['aplicables'] == 50
          and dos['completadas'] == 2
          and noaplica['aplicables'] == 49
          and noaplica['progreso'] > dos['progreso'])
    return {'ref': '04-onboarding-nuevo-empleado.xlsx:Checklist '
                   'Onboarding:{}'.format(c1),
            'tramos': tramos, 'tareas': total_tareas,
            'recien_descargado': fresco,
            'con_2_tareas_hechas': dos,
            'y_1_marcada_como_no_aplicable': noaplica,
            'ok': ok,
            'nota': 'la v1.1 barría `F7:F65` de corrido y se tragaba las 4 '
                    'cabeceras de sección, que valen literalmente «✓»: abría '
                    'en «4 de 51 completadas» y 8,51 % con el checklist en '
                    'blanco (DOM-01/TEC-02/TEC-13/COM-03)'}


def _d05_dias(carpeta):
    """§4 + RD-09/RD-10/RD-27/RT-03/RT-08 — el saldo de vacaciones.

    Lo que ya medía: agosto entero deja de contar como «1 día usado» y las
    fechas invertidas dejan de producir días negativos.

    Lo que se añade, todo cambiando entradas sobre MI copia:

      · **RD-09/RT-03**: tres periodos del MISMO empleado suman en el saldo.
        Con las 30 filas anteriores esa era la capacidad ENTERA de la hoja
        para toda la plantilla.
      · **RD-10**: una solicitud de OTRO ejercicio no descuenta del derecho de
        éste. Sin el filtro por año, reutilizar el libro dejaba a toda la
        plantilla en negativo.
      · **RD-27**: una baja a mitad de año prorratea por los dos extremos.
      · **RT-08**: un alta posterior al cierre del año da 0, no un derecho
        NEGATIVO, y lo dice en la columna de aviso.
    """
    p = os.path.join(carpeta, '05-planificacion-vacaciones.xlsx')
    if not os.path.isfile(p):
        return None
    cal, sol, sal = "'Calendario Anual'!", "'Solicitudes'!", \
                    "'Saldo Vacaciones'!"
    anio = ANIO_CALENDARIO
    salidas = [cal + '{}6'.format(COL_TOT)] + [sal + c + '5' for c in 'ABCEFGHI']

    def escena(sets, extra=()):
        xl = _xl(p)
        _calentar(xl, salidas + list(extra))
        for ref, val in sets:
            _calentar(xl, [ref])
            _sv(xl, ref, val)
        lect = dict((r.split('!')[-1], _ev(xl, r)) for r in salidas)
        for r in extra:
            lect[r] = _ev(xl, r)
        return lect

    def peticion(fila, ini, fin, estado='Aprobado', quien='Ana Ruiz'):
        return [(sol + 'A{}'.format(fila), quien),
                (sol + 'B{}'.format(fila), _serie(*ini)),
                (sol + 'C{}'.format(fila), _serie(*fin)),
                (sol + 'G{}'.format(fila), estado)]

    nombre = [(cal + 'A6', 'Ana Ruiz')]
    fresco = escena([])
    agosto = escena(nombre + peticion(5, (anio, 8, 1), (anio, 8, 30)),
                    extra=[sol + 'D5'])
    pendiente = escena(nombre + peticion(5, (anio, 8, 1), (anio, 8, 30),
                                         'Pendiente'))
    # RD-09/RT-03 · tres periodos de la misma persona
    tres = escena(nombre
                  + peticion(5, (anio, 8, 1), (anio, 8, 10))     # 10 días
                  + peticion(6, (anio, 12, 24), (anio, 12, 31))  # 8 días
                  + peticion(7, (anio, 4, 1), (anio, 4, 7)))     # 7 días
    # y una en la última fila de la hoja: el rango llega hasta ahí
    ultima = escena(nombre + peticion(SOL1, (anio, 5, 1), (anio, 5, 5)))
    # RD-10 · la misma petición, un año después
    otro_anio = escena(nombre + peticion(5, (anio + 1, 8, 1), (anio + 1, 8, 30)))
    # RD-27 · alta el 1 de enero y baja el 30 de junio → medio derecho
    baja = escena(nombre + [(sal + 'C5', _serie(anio, 6, 30))])
    # RT-08 · alta posterior al cierre del año
    imposible = escena(nombre + [(sal + 'B5', _serie(anio + 1, 3, 1))])
    invertidas = escena(nombre + peticion(5, (anio, 8, 30), (anio, 8, 1)),
                        extra=[sol + 'D5'])

    ok = (all(fresco[c] in ('', None) for c in ('E5', 'H5', 'BC6'))
          and agosto[sol + 'D5'] == 30 and agosto['E5'] == 30
          and agosto['F5'] == 30 and agosto['H5'] == 0
          and agosto['BC6'] == 30
          and pendiente['F5'] == 0 and pendiente['G5'] == 30
          and pendiente['H5'] == 0
          and tres['F5'] == 25 and tres['H5'] == 5
          and ultima['F5'] == 5
          and otro_anio['F5'] == 0 and otro_anio['H5'] == 30
          and _num(baja['E5']) and 14 < baja['E5'] < 16
          and imposible['E5'] == 0
          and 'posterior al cierre' in str(imposible['I5'])
          and 'invertidas' in str(invertidas[sol + 'D5']))
    return {'ref': '05-planificacion-vacaciones.xlsx:Saldo Vacaciones:'
                   'E5/F5/H5/I5',
            'hoja_recien_descargada': dict((k, str(v))
                                           for k, v in fresco.items()),
            'agosto entero aprobado':
                {'dias': agosto[sol + 'D5'], 'derecho': agosto['E5'],
                 'disfrutados': agosto['F5'], 'le_quedan': agosto['H5'],
                 'dias_usados_calendario': agosto['BC6']},
            'la misma solicitud en PENDIENTE':
                {'disfrutados': pendiente['F5'],
                 'pendientes': pendiente['G5'],
                 'le_quedan': pendiente['H5']},
            'TRES periodos de la misma persona (10 + 8 + 7 días) — RD-09':
                {'disfrutados': tres['F5'], 'le_quedan': tres['H5']},
            'una petición en la ÚLTIMA fila (%d) — RT-03' % SOL1:
                {'disfrutados': ultima['F5']},
            'la misma petición pero del año SIGUIENTE — RD-10':
                {'disfrutados': otro_anio['F5'],
                 'le_quedan': otro_anio['H5']},
            'baja el 30 de junio (medio año de alta) — RD-27':
                {'derecho': baja['E5']},
            'alta posterior al cierre del año — RT-08':
                {'derecho': imposible['E5'], 'aviso': imposible['I5']},
            'fechas al revés': str(invertidas[sol + 'D5']),
            'ok': ok,
            'nota': 'la v1.1 tenía una celda por MES: agosto entero contaba '
                    '1 día usado y 29 restantes, y las fechas invertidas '
                    'daban días NEGATIVOS que entraban en el saldo '
                    '(DOM-04/DOM-20/TEC-04/COM-06). Y la v2.0 dejaba la hoja '
                    'de peticiones en 30 filas para 30 personas (RD-09), '
                    'sumaba las de TODOS los años (RD-10), prorrateaba sólo '
                    'el alta (RD-27) y podía dar derecho negativo (RT-08)'}


def _d05_cobertura(carpeta):
    """§4 + RD-12/RD-13/RD-14/RT-11/RT-14 — la fila de cobertura y sus
    indicadores.

      · **RD-13**: un PERMISO RETRIBUIDO (PE) es una ausencia que hay que
        cubrir y no computaba: la fila sólo contaba V y B.
      · **RD-12**: la cobertura salía de las 1.590 celdas pintadas a mano
        mientras el saldo salía de las solicitudes; quien trabajaba con
        «Solicitudes» —que es lo que dicen las Instrucciones— dejaba la
        cobertura a cero.
      · **RT-11**: la temporada alta se evaluaba PRIMERO y con `>0`, así que
        una sola ausencia pintaba el bloqueante ⛔ y el «⚠ EXCESO» —la única
        alerta que compara contra el máximo del cliente— no salía nunca.
      · **RD-14/RT-14**: el bloque de personal mínimo por turno no lo leía
        ninguna fórmula.
    """
    p = os.path.join(carpeta, '05-planificacion-vacaciones.xlsx')
    if not os.path.isfile(p):
        return None
    cal, cob, sol = "'Calendario Anual'!", "'Cobertura'!", "'Solicitudes'!"
    # semana 1 (normal) y la primera precargada como «Alta»
    wb = openpyxl.load_workbook(p)
    hoja = wb['Calendario Anual']
    sem_alta = None
    for n in range(1, SEMANAS + 1):
        col = get_column_letter(1 + n)
        if hoja['{}{}'.format(col, F_TEMP)].value == 'Alta':
            sem_alta = col
            break
    sem_normal = 'B'
    salidas = [cal + '{}{}'.format(c, f)
               for c in (sem_normal, sem_alta)
               for f in (F_TEMP, F_AUS, F_SOL, F_MAX, F_ALERTA)]
    salidas += [cob + 'B{}'.format(r) for r in (6, 7, 8, 44, 45, 46, 47, 48)]

    def escena(sets):
        xl = _xl(p)
        _calentar(xl, salidas)
        for ref, val in sets:
            _calentar(xl, [ref])
            _sv(xl, ref, val)
        return dict((r.split('!')[-1], _ev(xl, r)) for r in salidas)

    plantilla = [(cal + 'A{}'.format(EMP0 + i), 'Empleado {}'.format(i + 1))
                 for i in range(10)]

    def marca(col, codigos):
        return [(cal + '{}{}'.format(col, EMP0 + i), c)
                for i, c in enumerate(codigos)]

    fresco = escena([])
    tres = escena(plantilla + marca(sem_normal, ['V', 'V', 'V']))
    # RD-13 · dos vacaciones y dos permisos retribuidos
    permisos = escena(plantilla + marca(sem_normal, ['V', 'V', 'PE', 'PE']))
    # RT-11 · el exceso manda sobre la temporada
    exceso = escena(plantilla + marca(sem_normal, ['V'] * 6))
    exceso_alta = escena(plantilla + marca(sem_alta, ['V'] * 6))
    una_en_alta = escena(plantilla + marca(sem_alta, ['V']))
    # RD-12 · sin pintar NADA en la rejilla, sólo con solicitudes aprobadas
    lunes = LUNES_SEMANA_1
    solicitudes = escena(
        plantilla
        + [(sol + 'A5', 'Empleado 1'),
           (sol + 'B5', _serie(lunes.year, lunes.month, lunes.day)),
           (sol + 'C5', _serie((lunes + datetime.timedelta(days=4)).year,
                               (lunes + datetime.timedelta(days=4)).month,
                               (lunes + datetime.timedelta(days=4)).day)),
           (sol + 'G5', 'Aprobado')])
    # RD-14/RT-14 · el mínimo por turno decide un indicador
    sin_cubrir = escena(plantilla + marca(sem_normal, ['V'] * 7))

    ok = (fresco['{}{}'.format(sem_normal, F_MAX)] == 0
          and fresco['{}{}'.format(sem_normal, F_ALERTA)] in ('', None)
          and fresco['B47'] == 0 and fresco['B48'] == 0
          and tres['{}{}'.format(sem_normal, F_AUS)] == 3
          and tres['{}{}'.format(sem_normal, F_MAX)] == 3
          and tres['B44'] == 10 and tres['B45'] == 3
          and tres['{}{}'.format(sem_normal, F_ALERTA)] in ('', None)
          and permisos['{}{}'.format(sem_normal, F_AUS)] == 4
          and 'EXCESO' in str(exceso['{}{}'.format(sem_normal, F_ALERTA)])
          and 'TEMP. ALTA' not in str(
              exceso['{}{}'.format(sem_normal, F_ALERTA)])
          and 'EXCESO en TEMP. ALTA' in str(
              exceso_alta['{}{}'.format(sem_alta, F_ALERTA)])
          and 'TEMP. ALTA' in str(
              una_en_alta['{}{}'.format(sem_alta, F_ALERTA)])
          and 'EXCESO' not in str(
              una_en_alta['{}{}'.format(sem_alta, F_ALERTA)])
          and solicitudes['{}{}'.format(sem_normal, F_AUS)] == 0
          and solicitudes['{}{}'.format(sem_normal, F_SOL)] == 1
          and solicitudes['{}{}'.format(sem_normal, F_MAX)] == 1
          and sin_cubrir['B48'] >= 1
          and tres['B48'] == 0
          and permisos['B6'] == 4 and permisos['B7'] == 5
          and permisos['B8'] == 2)
    return {'ref': '05-planificacion-vacaciones.xlsx:Calendario '
                   'Anual:B{}:B{}'.format(F_AUS, F_ALERTA),
            'hoja_recien_descargada':
                {'ausencias': fresco['{}{}'.format(sem_normal, F_MAX)],
                 'alerta': str(fresco['{}{}'.format(sem_normal, F_ALERTA)]),
                 'semanas_sin_cubrir_B48': fresco['B48']},
            'tres ausentes la semana 1 (máximo 4)':
                {'rejilla': tres['{}{}'.format(sem_normal, F_AUS)],
                 'la_mayor': tres['{}{}'.format(sem_normal, F_MAX)],
                 'plantilla': tres['B44'], 'pico': tres['B45'],
                 'alerta': str(tres['{}{}'.format(sem_normal, F_ALERTA)])},
            'dos V y dos PE — el permiso ahora cuenta (RD-13)':
                {'ausencias': permisos['{}{}'.format(sem_normal, F_AUS)]},
            'seis ausentes en semana NORMAL (RT-11)':
                {'alerta': str(exceso['{}{}'.format(sem_normal, F_ALERTA)])},
            'seis ausentes en TEMPORADA ALTA (RT-11)':
                {'alerta': str(
                    exceso_alta['{}{}'.format(sem_alta, F_ALERTA)])},
            'UNA ausencia en temporada alta (RT-11)':
                {'alerta': str(
                    una_en_alta['{}{}'.format(sem_alta, F_ALERTA)])},
            'sin pintar nada, sólo una solicitud APROBADA (RD-12)':
                {'rejilla': solicitudes['{}{}'.format(sem_normal, F_AUS)],
                 'por_solicitudes':
                     solicitudes['{}{}'.format(sem_normal, F_SOL)],
                 'la_mayor': solicitudes['{}{}'.format(sem_normal, F_MAX)]},
            'mínimo por turno (M · T · N)': {'M': permisos['B6'],
                                             'T': permisos['B7'],
                                             'N': permisos['B8']},
            'semanas que NO cubren el turno más exigente (RD-14/RT-14)':
                {'con 3 ausentes': tres['B48'],
                 'con 7 ausentes de 10': sin_cubrir['B48']},
            'ok': ok,
            'nota': "la hoja 'Cobertura' de la v1.1 eran dos cabeceras y cero "
                    'datos (COM-20); la v2.0 la llenó pero la dejó sin '
                    'conectar (RD-14/RT-14), contaba sólo V y B (RD-13), no '
                    'se enteraba de las solicitudes aprobadas (RD-12) y '
                    'tapaba el exceso con la temporada alta (RT-11)'}


def demos(carpeta, origen):
    fuera = {}
    for nombre, fn in (('grupo_b_03_coste_por_empleado', _d03_coste),
                       ('grupo_b_03_semaforo_por_tipo', _d03_semaforo),
                       ('grupo_b_03_fte_por_servicio', _d03_fte),
                       ('grupo_b_04_progreso_onboarding', _d04_progreso),
                       ('grupo_b_05_dias_y_saldo', _d05_dias),
                       ('grupo_b_05_cobertura', _d05_cobertura)):
        try:
            r = fn(carpeta)
        except Exception as e:                              # noqa: BLE001
            r = {'ok': False, 'error': '{}: {}'.format(type(e).__name__, e)}
        if r is not None:
            fuera[nombre] = r
    return fuera
