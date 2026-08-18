#!/usr/bin/env python3
"""
Kit de Tareas Pastelería / Obrador — revisión v1.1 (2026-08-18), post-proceso DETERMINISTA
sobre los 11 .xlsx LIVE (sin regenerar): metadata de marca, línea de versión, bio anclada,
configuración de impresión A4, casilla de completado unificada (bug L1-02), correcciones
técnicas/calendario objetivas (auditoría v1, lente obrador) y, al final, cache de valores
(inject_cache.py). Idempotente: se puede volver a ejecutar.

Uso: python3 scripts/productos-digitales/kit-pasteleria-v1_1-postprocess.py
"""
import os, re, subprocess, sys
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.page import PageMargins

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl', 'kit-tareas-pasteleria')
VERSION_LINE = 'Versión 1.1 · agosto 2026 · aichef.pro/kit-tareas-pasteleria · info@aichef.pro'
BIO_OLD = '29 años de experiencia en alta hostelería · 15 años de consultoría gastronómica'
BIO_NEW = 'Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, en cocina desde los 17 años · johnguerrero.es'
MARK_LINE = 'Marca con ✓ en la columna «Hecha» (desplegable): es la que cuenta el total de tareas completadas.'

# Correcciones de texto exactas (fichero → {texto viejo: texto nuevo}); solo sustituciones 1:1 en la misma celda.
TEXT_FIXES = {
    '02-partidas-cocina.xlsx': {
        'Amasar en planetaria (1ª velocidad 4 min, 2ª velocidad 8-10 min)':
            'Amasar corto en planetaria (1ª velocidad 3-4 min, 2ª velocidad 2-3 min): el gluten se termina de desarrollar en el laminado',
        'Control de temperatura de masa al salir (24-26 °C ideal)':
            'Control de temperatura de masa al salir (22-24 °C; 20-22 °C si va directa a bloque de frío)',
        'Congelar entremets montados (mínimo 4h antes de glasear)':
            'Congelar entremets montados para el glaseado de mañana (mín. 4 h en abatidor a −18 °C o 12 h en congelador)',
        'Desmoldar entremets congelados':
            'Desmoldar los entremets congelados el día anterior',
        'Montar tartas por encargo (según ficha de pedido)':
            'Montar tartas por encargo (según la ficha de encargo de cada pedido)',
    },
    '06-eventos-festivos.xlsx': {
        'Hornear roscones la madrugada del 5 al 6':
            'Hornear roscones: 1ª hornada la madrugada del 4 al 5 (60-70 % del volumen, venta de la tarde del 5) y 2ª hornada del 5 al 6',
        '05-06 Ene': '04-06 Ene',
        'Elaborar torrijas (remojo, fritura, almíbar/canela)':
            'Elaborar torrijas a diario (remojo, fritura, almíbar/canela) — pico Jueves y Viernes Santo',
        'Jue-Vie Santo': 'V. Dolores → D. Resurrección',
        'Elaborar monas de Pascua (bizcocho + decoración chocolate)':
            'Elaborar y montar monas de Pascua (bizcocho + decoración chocolate) para recogida el Domingo de Resurrección y el Lunes de Pascua',
        'Lun-Mié': 'Jue-Sáb Santo',
    },
    'BONUS-02-calendario-anual-tareas.xlsx': {
        '25 Nov': 'Último viernes de nov. (variable)',
    },
}


def set_metadata(wb, fname):
    num = fname.split('-')[0]
    nice = {
        '01': 'Apertura y Cierre', '02': 'Partidas de Producción', '03': 'Tareas del Manager',
        '04': 'Tareas por Perfil', '05': 'Tareas Semanales y Mensuales', '06': 'Eventos y Festivos',
        '07': 'Plantilla Personalizable', '08': 'Apertura y Cierre del Negocio', '09': 'Apertura y Cierre de Caja',
        'BONUS': 'Bonus',
    }.get(num, fname)
    if fname.startswith('BONUS-01'):
        nice = 'Bonus 01 — Briefing de Servicio'
    if fname.startswith('BONUS-02'):
        nice = 'Bonus 02 — Calendario Anual'
    p = wb.properties
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'
    p.title = f'{num} — {nice} · Kit de Tareas Pastelería / Obrador' if not fname.startswith('BONUS') else f'{nice} · Kit de Tareas Pastelería / Obrador'
    p.subject = 'Kit de Tareas Recurrentes · Pastelería / Obrador · v1.1'
    p.keywords = 'pastelería, obrador, checklist, tareas, AI Chef Pro'
    p.description = 'aichef.pro/kit-tareas-pasteleria'
    p.category = 'AI Chef Pro · Productos digitales'


def print_setup(ws, header_row, landscape=True):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59, header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = f'{header_row}:{header_row}'
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def find_header_row(ws):
    """Fila cuya col. B (o A) es 'Tarea' / '#'. Devuelve None si no hay tabla."""
    for r in range(1, 8):
        b = ws.cell(row=r, column=2).value
        a = ws.cell(row=r, column=1).value
        if b == 'Tarea' or a in ('#', 'Fecha', 'Denominación'):
            return r
    return None


def fix_checkbox_family(ws, header_row):
    """Familia «▸» (01-07, BONUS-01): col. A tiene ☐ y el contador cuenta la col. F.
    Unificar: A pasa a numeración de tarea; F ya tiene validación «✓,—»."""
    if ws.cell(row=header_row, column=1).value != '☐':
        return 0
    ws.cell(row=header_row, column=1).value = 'Nº'
    n = 0
    for r in range(header_row + 1, ws.max_row + 1):
        c = ws.cell(row=r, column=1)
        if c.value == '☐':
            n += 1
            c.value = n
            c.font = Font(name='Calibri', size=10, color='666666')
            c.alignment = Alignment(horizontal='center', vertical='center')
    return n


def add_line_instructions(ws, text, bold=False, size=9, color='999999'):
    col = 2 if ws.cell(row=2, column=2).value else 1
    last = ws.max_row
    # evitar duplicar en re-ejecuciones
    for r in range(1, last + 1):
        if ws.cell(row=r, column=col).value == text:
            return
    ws.cell(row=last + 2, column=col, value=text).font = Font(name='Calibri', size=size, bold=bold, color=color)


def process(fname):
    path = os.path.join(DL, fname)
    wb = openpyxl.load_workbook(path)
    set_metadata(wb, fname)
    fixes = TEXT_FIXES.get(fname, {})
    applied = 0
    marks = 0
    for ws in wb.worksheets:
        # texto
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    if c.value in fixes:
                        c.value = fixes[c.value]; applied += 1
                    elif c.value == BIO_OLD:
                        c.value = BIO_NEW; applied += 1
        if ws.title == 'Instrucciones':
            fam = ws.cell(row=2, column=2).value is not None  # familia «▸» escribe en col. B
            if fam and fname not in ('BONUS-02-calendario-anual-tareas.xlsx',):
                add_line_instructions(ws, MARK_LINE, bold=True, size=11, color='333333')
            add_line_instructions(ws, VERSION_LINE)
            ws.page_setup.paperSize = 9
            ws.page_setup.orientation = 'portrait'
            continue
        hr = find_header_row(ws)
        if hr:
            marks += fix_checkbox_family(ws, hr)
        landscape = ws.max_column >= 6
        print_setup(ws, hr, landscape)
    wb.save(path)
    return applied, marks


def verify(fname):
    path = os.path.join(DL, fname)
    wb = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path)
    nform = 0; nocache = 0; nonlatin = 0
    rx = re.compile('[぀-ヿ㐀-䶿一-鿿가-힯Ѐ-ӿ؀-ۿ֐-׿฀-๿]')
    for ws, wsf in zip(wb.worksheets, wbf.worksheets):
        for row in wsf.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    nform += 1
                    if ws[c.coordinate].value is None:
                        nocache += 1
                if isinstance(c.value, str) and rx.search(c.value):
                    nonlatin += 1
    bio_old = any(isinstance(c.value, str) and ('29 años' in c.value or '15 años' in c.value)
                  for ws in wbf.worksheets for row in ws.iter_rows() for c in row)
    return dict(creator=wbf.properties.creator, formulas=nform, sin_cache=nocache, no_latino=nonlatin, bio_vieja=bio_old,
                hojas=len(wbf.worksheets))


if __name__ == '__main__':
    files = sorted(f for f in os.listdir(DL) if f.endswith('.xlsx'))
    print(f'{len(files)} ficheros en {DL}')
    for f in files:
        a, m = process(f)
        print(f'  {f}: fixes_texto={a} casillas_renumeradas={m}')
    # cache al FINAL (cualquier save posterior lo borra)
    subprocess.run([sys.executable, os.path.join(ROOT, 'scripts', 'productos-digitales', 'inject_cache.py')] + [os.path.join(DL, f) for f in files], check=True)
    ok = True
    for f in files:
        v = verify(f)
        flag = v['sin_cache'] == 0 and v['no_latino'] == 0 and not v['bio_vieja'] and v['creator'] == 'AI Chef Pro'
        ok &= flag
        print(f"  {'OK ' if flag else 'FAIL'} {f}: {v}")
    print('TODO OK' if ok else 'HAY FALLOS')
    sys.exit(0 if ok else 1)
