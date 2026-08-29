#!/usr/bin/env python3
"""
censo-entregables.py — Gate DETERMINISTA de los entregables (xlsx/docx/pdf) de
productos digitales en `astro-site/public/dl/**`. NO modifica nada: solo lee y
mide. Es la vara de medir de la Fase A (ver
scripts/productos-digitales/FASE-A-SPEC-postprocess-transversal.md §1) y de
`postprocess-transversal.py`: antes/después de sanear, se corre este censo y se
compara.

Qué mide cada campo y por qué (para no reinventar la rueda leyendo el código):

  Por PRODUCTO (= nombre de carpeta bajo dl/). Los ficheros sueltos de la raíz
  de dl/ (39 huérfanos: ap-*.xlsx, ke-*.xlsx, b1-*.docx, b23-*.xlsx…) se
  etiquetan prod='dl' — se listan y entran en --json, pero NUNCA cuentan para
  --fail: no pertenecen a ningún producto vivo y no los toca la Fase A.

  Por XLSX (openpyxl, dos aperturas: normal + data_only=True):
    creator        wb.properties.creator — debe ser 'AI Chef Pro' tras Fase A.
    title          wb.properties.title — para verificar el patrón
                   "<Título legible> · <Nombre corto>" a ojo, no lo valida el gate.
    hojas          nº de hojas del libro.
    form           nº de celdas cuyo valor (wb SIN data_only) empieza por '='.
                   Es el total de fórmulas, no un defecto en sí.
    nocache_total  de esas fórmulas, cuántas devuelven None al abrir CON
                   data_only=True — es decir, sin valor cacheado en el XML.
                   openpyxl escribe el cache vacío (<v/>) y algunos visores
                   (Quick Look, móvil, Excel en cálculo manual) muestran la
                   celda en blanco hasta que alguien fuerza el recálculo.
    nocache_vacio  subconjunto de nocache_total cuya fórmula CONTIENE una
                   cadena vacía ("") en su cuerpo — típicamente
                   IF(x="","",...). openpyxl SÍ cachea correctamente esas
                   como cadena vacía, y data_only las devuelve como None
                   (no hay forma de distinguir "" de None en el XML), así
                   que un None aquí es el comportamiento ESPERADO, no un
                   fallo de inject_cache.py. Por eso el spec dice
                   explícitamente que nocache_vacio NO es un defecto.
    nocache_real   = nocache_total − nocache_vacio. Este es el que SÍ importa:
                   fórmulas que deberían tener valor cacheado y no lo tienen
                   (inject_cache.py no se corrió, o se corrió y luego alguien
                   volvió a guardar con openpyxl, que borra el cache al vuelo).
                   Es el campo que activa --fail.
    nonlat         nº de celdas de texto con al menos un carácter no latino
                   (CJK, cirílico, hangul, árabe, hebreo, tailandés) — la
                   familia de bugs de inyección documentada en CLAUDE.md
                   (Amanita phalloides en fermentus-ai, etc.). Cualquier valor
                   >0 es defecto: en un producto de pastelería/hostelería en
                   español no hay motivo legítimo para que aparezcan.
    bio_vieja      True si alguna celda casa con el patrón de la bio antigua
                   (29/15 años…) tal como se escribía antes del anclaje a
                   John Guerrero (2026-08-18/22). El patrón está afinado para
                   NO casar la bio nueva ("desde los 17 años", "consultor
                   gastronómico desde 2010") — si casara, cualquier fichero ya
                   corregido volvería a salir en rojo para siempre.
    box_colA       nº de celdas con valor exacto '☐' en la COLUMNA A. Antes de
                   la Fase A esto es casi siempre un defecto (checkbox muerta,
                   sin desplegable, que no suma al contador — bug L1-02). PERO
                   el patrón P3a de la spec (panadería-obrador, checklists de
                   apertura…) usa la ☐ de la columna A como LA marca legítima,
                   y tras normalizarla lleva validación de datos (lista) en
                   esa misma celda. Por eso también se mide:
    box_colA_dv    de esas ☐ en columna A, cuántas tienen una validación de
                   datos tipo 'list' aplicada a la celda — es la señal
                   objetiva de "☐ con desplegable" (P3a ya saneado) frente a
                   "☐ muerta" (defecto real, sin desplegable). El campo que
                   activa --fail es box_colA_muerta = box_colA − box_colA_dv.
    noprint        nº de hojas cuyo page_setup.paperSize != 9 (9 = A4). Antes
                   de la Fase A casi todas las hojas nuevas no tienen
                   paperSize fijado (None) y algunos visores imprimen en
                   Letter o con márgenes por defecto.
    autoria        True si alguna celda nombra a John Guerrero (la firma de
                   autoría que la Fase A ancló). INFORMATIVO, no activa --fail:
                   sirve para medir la cobertura de la bio, que en kit-tareas
                   sólo estaba en 2 de los 11 ficheros porque la Fase A
                   SUSTITUYE bios existentes y nunca las AÑADE (DOM-30 de la
                   R1). Se mide sobre todo el libro, no sólo sobre
                   'Instrucciones': hay generadores que la ponen en el pie.
    version_line   True si alguna celda empieza por 'Versión ' o 'Version ' —
                   para detectar de un vistazo si el fichero ya lleva la línea
                   de versión/marca en Instrucciones. Informativo, no falla el
                   gate por sí solo (dentro de las hojas de Instrucciones que
                   no existen en todos los ficheros).
    cb_brand       nº de celdas con 'ChefBusiness' en el texto — informativo:
                   son los 7 productos de origen ChefBusiness cuya marca NO se
                   toca sin el OK de John (spec §3). No activa --fail.
    empty_str      nº de celdas cuyo valor es exactamente '' (cadena vacía,
                   no None). Es XML no conforme de cara a algunos parsers y
                   el post-proceso las normaliza a None (spec §2.6).
    hoja_larga     lista de títulos de hoja de más de 31 caracteres, que es el
                   máximo de Excel: al abrir, Excel los recorta y cualquier
                   referencia entre hojas que use el nombre largo se rompe.
                   INFORMATIVO — no activa --fail.

  Por DOCX (python-docx):
    parrafos       nº de párrafos no vacíos — orientativo, no defecto.
    bio_vieja      mismo patrón que en xlsx, sobre el texto de cada párrafo.
    autoria        mismo patrón que en xlsx, sobre el texto de cada párrafo.
    nonlat         mismo patrón que en xlsx, sobre el texto de cada párrafo.

  Por PDF: si hay `pdftotext` en el PATH se extrae texto y se mide nonlat
  sobre él; si no, el campo queda como 'skip' (no se inventa un resultado).

Uso:
  python3 censo-entregables.py [--only <producto>] [--json salida.json]
                                [--fail] [--quiet]

  --only <producto>  Solo censa esa carpeta de dl/ (p. ej. kit-tareas-pasteleria).
                     También acepta una RUTA de carpeta existente (con
                     separador de directorios): la recorre tal cual y usa su
                     basename como nombre de producto — es como se censa la
                     copia de `--dry-run` en el scratchpad sin tocar /dl/.
  --json <path>       Vuelca la lista COMPLETA (un dict por fichero, misma forma
                       que censo-entregables-2026-08-22.json más los campos
                       nuevos de esta versión) en ese path, en JSON compacto.
  --fail               Exit 1 si algún xlsx de un PRODUCTO (no de 'dl' huérfano)
                       incumple alguno de: nocache_real>0, box_colA_muerta>0,
                       bio_vieja, noprint>0, creator != 'AI Chef Pro', nonlat>0,
                       empty_str>0. Sin --fail el exit code es siempre 0
                       (modo informe).
  --quiet              No imprime la tabla por producto; solo el resumen final
                       (y el --json si se pidió). Útil para correrlo en bucle.

Tiempo de referencia: ~2 min para los ~645 ficheros de dl/ (dos aperturas de
openpyxl por xlsx: normal y data_only). Corre EN SERIE — no paralelizar: es
carga de CPU sostenida y el Mac tiene el límite térmico de ~65 °C de CLAUDE.md.
"""
import argparse
import json
import os
import re
import shutil
import subprocess
import sys

import openpyxl

try:
    import docx  # python-docx
except ImportError:
    docx = None

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl')

RX_NONLAT = re.compile(
    '[぀-ヿ㐀-䶿一-鿿가-힯Ѐ-ӿ؀-ۿ֐-׿฀-๿]'
)
RX_BIO_VIEJA = re.compile(
    r'29 a[ñn]os|15 a[ñn]os|a[ñn]os de experiencia|15\+ a[ñn]os|29\+ a[ñn]os'
)
#: Firma de autoría anclada (informativa). Basta el nombre: la redacción larga
#: («chef y consultor gastronómico desde 2010…») convive con variantes cortas
#: en algunos pies, y lo que se quiere medir es si el fichero atribuye o no.
RX_AUTORIA = re.compile(r'John\s+Guerrero', re.I)

#: Falsos positivos REALES y documentados de RX_BIO_VIEJA: no son bios, son
#: contenido de negocio que casualmente casa el patrón — la vida útil de un
#: horno Josper, en
#: `kit-tareas-asador/07-semanales-mensuales.xlsx:Instrucciones:B6`.
#: ⚠️ MISMO LITERAL que `FALSOS_POSITIVOS_BIO` de
#: `postprocess-transversal.py`: si se toca uno hay que tocar el otro, o el gate
#: y el post-proceso dejarán de coincidir y el saneamiento nunca dará verde.
FALSOS_POSITIVOS_BIO = frozenset({
    '▸ El mantenimiento del Josper es clave: un horno bien cuidado dura 15+ anos',
    # CB-E1 (2026-08-29) — la MISMA frase con la ñ puesta. `motor.ortografia`
    # corrige «anos» → «años» en los siete kits de ChefBusiness, así que la
    # entrada de arriba deja de casar y el gate marcaba `bio_vieja` en
    # `kit-tareas-asador/07-semanales-mensuales.xlsx` — una lista blanca por
    # LITERAL EXACTO invalidada por una normalización posterior. Se conservan
    # las dos formas: la vieja mientras el fichero publicado siga sin la ñ.
    '▸ El mantenimiento del Josper es clave: un horno bien cuidado dura 15+ años',
})

#: Funciones de Excel que pycel NO implementa: la celda se queda sin cache por
#: limitación del evaluador, no porque el fichero esté mal (Excel la calcula al
#: abrir). Se cuentan aparte como `nocache_pycel` y NO son defecto.
#: ⚠️ MISMO LITERAL que `FUNCIONES_SIN_PYCEL` de `postprocess-transversal.py`
#: (misma regla de sincronización que FALSOS_POSITIVOS_BIO). Caso medido:
#: `kit-plan-financiero/07-informe-viabilidad-bancos.xlsx:Proyecciones:B21`, IRR.
FUNCIONES_SIN_PYCEL = ('IRR', 'XIRR', 'PMT', 'COUNTA')
RX_SIN_PYCEL = re.compile(r'\b(' + '|'.join(FUNCIONES_SIN_PYCEL) + r')\s*\(', re.I)


def _es_nonlat(v):
    return isinstance(v, str) and bool(RX_NONLAT.search(v))


def _es_bio_vieja(v):
    if not isinstance(v, str) or v in FALSOS_POSITIVOS_BIO:
        return False
    return bool(RX_BIO_VIEJA.search(v))


def _dv_lista_en(ws, coord):
    """True si la celda `coord` tiene una validación de datos tipo 'list'."""
    for dv in ws.data_validations.dataValidation:
        if dv.type == 'list' and coord in dv.sqref:
            return True
    return False


def censar_xlsx(path, prod, fname):
    r = dict(prod=prod, f=fname, tipo='xlsx')
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        wbv = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        r['error'] = f'{type(e).__name__}: {e}'
        return r

    p = wb.properties
    r['creator'] = p.creator
    r['title'] = p.title
    r['hojas'] = len(wb.worksheets)
    # Informativo (no cuenta para --fail): Excel corta las pestañas a 31
    # caracteres, así que un título más largo se renombra solo al abrir y las
    # referencias entre hojas que lo mencionen dejan de resolver.
    r['hoja_larga'] = [t for t in wb.sheetnames if len(t) > 31]

    form = 0
    nocache_total = 0
    nocache_vacio = 0
    nocache_pycel = 0
    nonlat = 0
    bio_vieja = False
    autoria = False
    box_colA = 0
    box_colA_dv = 0
    version_line = False
    cb_brand = 0
    empty_str = 0
    noprint = 0

    for ws, wsv in zip(wb.worksheets, wbv.worksheets):
        # A4 COMPLETO: paperSize 9 + ajuste al ancho + pie (un paperSize=9 a secas
        # es el mínimo del generador: 147 hojas así se colaron el 2026-08-22)
        fit = (ws.sheet_properties.pageSetUpPr.fitToPage
               if ws.sheet_properties.pageSetUpPr else None)
        pie = ws.oddFooter.center.text if ws.oddFooter and ws.oddFooter.center else None
        es_texto = ws.title in ('Instrucciones', 'Índice', 'Indice')
        if ws.page_setup.paperSize != 9 or (not es_texto and (not fit or not pie)):
            noprint += 1
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                # fórmula = tipo de celda 'f', no «texto que empieza por =»: las
                # etiquetas «= Margen bruto» convertidas a texto no son fórmulas
                if c.data_type == 'f' and isinstance(v, str):
                    form += 1
                    if wsv[c.coordinate].value is None:
                        nocache_total += 1
                        if '""' in v:
                            nocache_vacio += 1
                        elif RX_SIN_PYCEL.search(v):
                            nocache_pycel += 1
                if _es_nonlat(v):
                    nonlat += 1
                if _es_bio_vieja(v):
                    bio_vieja = True
                if isinstance(v, str) and RX_AUTORIA.search(v):
                    autoria = True
                if isinstance(v, str) and (v.startswith('Versión ') or v.startswith('Version ')):
                    version_line = True
                if isinstance(v, str) and 'ChefBusiness' in v:
                    cb_brand += 1
                if v == '':
                    empty_str += 1
                if c.column == 1 and v == '☐':
                    box_colA += 1
                    if _dv_lista_en(ws, c.coordinate):
                        box_colA_dv += 1

    r.update(
        form=form,
        nocache_total=nocache_total,
        nocache_vacio=nocache_vacio,
        nocache_pycel=nocache_pycel,
        nocache_real=nocache_total - nocache_vacio - nocache_pycel,
        nonlat=nonlat,
        bio_vieja=bio_vieja,
        autoria=autoria,
        box_colA=box_colA,
        box_colA_dv=box_colA_dv,
        box_colA_muerta=box_colA - box_colA_dv,
        noprint=noprint,
        version_line=version_line,
        cb_brand=cb_brand,
        empty_str=empty_str,
    )
    return r


def censar_docx(path, prod, fname):
    r = dict(prod=prod, f=fname, tipo='docx')
    if docx is None:
        r['error'] = 'python-docx no instalado'
        return r
    try:
        d = docx.Document(path)
    except Exception as e:
        r['error'] = f'{type(e).__name__}: {e}'
        return r
    parrafos = 0
    nonlat = 0
    bio_vieja = False
    autoria = False
    for para in d.paragraphs:
        t = para.text
        if t.strip():
            parrafos += 1
        if _es_nonlat(t):
            nonlat += 1
        if _es_bio_vieja(t):
            bio_vieja = True
        if RX_AUTORIA.search(t):
            autoria = True
    r.update(parrafos=parrafos, nonlat=nonlat, bio_vieja=bio_vieja,
             autoria=autoria)
    return r


def censar_pdf(path, prod, fname):
    r = dict(prod=prod, f=fname, tipo='pdf')
    if shutil.which('pdftotext') is None:
        r['skip'] = True
        return r
    try:
        out = subprocess.run(['pdftotext', path, '-'], capture_output=True, text=True, timeout=60)
        texto = out.stdout
    except Exception as e:
        r['error'] = f'{type(e).__name__}: {e}'
        return r
    r.update(nonlat=1 if _es_nonlat(texto) else 0, bio_vieja=_es_bio_vieja(texto))
    return r


def _recorrer_carpeta(ruta, prod):
    """(prod, fname, path) de todo xlsx/docx/pdf que cuelgue de `ruta`."""
    items = []
    for dirpath, _dirnames, filenames in os.walk(ruta):
        for fn in sorted(filenames):
            if fn.lower().endswith(('.xlsx', '.docx', '.pdf')):
                items.append((prod, fn, os.path.join(dirpath, fn)))
    return items


def recorrer(only=None):
    """Genera (prod, fname, path) para cada xlsx/docx/pdf bajo dl/. Carpeta =
    producto; ficheros sueltos de la raíz de dl/ -> prod='dl'.

    `only` acepta un nombre de producto O una RUTA de carpeta existente
    (decisión SPC-05): así se puede censar la copia que deja
    `postprocess-transversal.py --dry-run` en el scratchpad sin tener que
    tocar `/dl/`. En ese caso el producto es el `basename` de la ruta, que es
    justo el nombre que tiene la carpeta copiada.
    """
    if only and (os.sep in only or only in ('.', '..')):
        ruta = os.path.abspath(only).rstrip(os.sep)
        if not os.path.isdir(ruta):
            return []
        return _recorrer_carpeta(ruta, os.path.basename(ruta))

    items = []
    for entry in sorted(os.listdir(DL)):
        full = os.path.join(DL, entry)
        if os.path.isdir(full):
            if entry.startswith('.'):
                continue
            if only and entry != only:
                continue
            items.extend(_recorrer_carpeta(full, entry))
        elif os.path.isfile(full):
            if only and only != 'dl':
                continue
            if entry.lower().endswith(('.xlsx', '.docx', '.pdf')):
                items.append(('dl', entry, full))
    return items


def censar_uno(prod, fname, path):
    ext = fname.lower().rsplit('.', 1)[-1]
    if ext == 'xlsx':
        return censar_xlsx(path, prod, fname)
    if ext == 'docx':
        return censar_docx(path, prod, fname)
    if ext == 'pdf':
        return censar_pdf(path, prod, fname)
    raise ValueError(fname)


DEFECTOS_XLSX = (
    'nocache_real', 'box_colA_muerta', 'noprint', 'nonlat', 'empty_str',
)


def es_defectuoso_xlsx(r):
    if r.get('tipo') != 'xlsx' or 'error' in r:
        return bool(r.get('error'))
    if r.get('creator') != 'AI Chef Pro':
        return True
    if r.get('bio_vieja'):
        return True
    return any(r.get(k, 0) > 0 for k in DEFECTOS_XLSX)


def tabla_por_producto(rows):
    """Agrupa por producto (excluye 'dl' huérfano) y arma las columnas de la
    tabla de la spec: n, form, f_sincache_real, box, bio, noA4, creator≠AICP."""
    prods = {}
    for r in rows:
        if r['prod'] == 'dl':
            continue
        agg = prods.setdefault(r['prod'], dict(
            n=0, form=0, f_sincache_real=0, box=0, bio=0, noA4=0, creator_mal=0,
            nonlat=0, empty_str=0, errores=0,
        ))
        agg['n'] += 1
        if r.get('tipo') != 'xlsx':
            continue
        if 'error' in r:
            agg['errores'] += 1
            continue
        agg['form'] += r.get('form', 0)
        if r.get('nocache_real', 0) > 0:
            agg['f_sincache_real'] += 1
        if r.get('box_colA_muerta', 0) > 0:
            agg['box'] += 1
        if r.get('bio_vieja'):
            agg['bio'] += 1
        if r.get('noprint', 0) > 0:
            agg['noA4'] += 1
        if r.get('creator') != 'AI Chef Pro':
            agg['creator_mal'] += 1
        if r.get('nonlat', 0) > 0:
            agg['nonlat'] += 1
        if r.get('empty_str', 0) > 0:
            agg['empty_str'] += 1
    return prods


def imprimir_tabla(prods):
    cab = ('producto', 'n', 'form', 'f_sinCache', 'box', 'bio', 'noA4', 'creator≠AICP', 'nonlat', 'emptyStr', 'err')
    ancho = [max(len(cab[i]), 10) for i in range(len(cab))]
    print('  '.join(c.ljust(ancho[i]) for i, c in enumerate(cab)))
    tot = dict(n=0, form=0, f_sincache_real=0, box=0, bio=0, noA4=0, creator_mal=0, nonlat=0, empty_str=0, errores=0)
    for prod in sorted(prods):
        a = prods[prod]
        fila = (prod, a['n'], a['form'], a['f_sincache_real'], a['box'], a['bio'], a['noA4'],
                a['creator_mal'], a['nonlat'], a['empty_str'], a['errores'])
        print('  '.join(str(v).ljust(ancho[i]) for i, v in enumerate(fila)))
        for k in tot:
            tot[k] += a[k]
    print('-' * sum(ancho))
    print(f"TOTAL: n={tot['n']} form={tot['form']} f_sinCache={tot['f_sincache_real']} "
          f"box={tot['box']} bio={tot['bio']} noA4={tot['noA4']} "
          f"creator≠AICP={tot['creator_mal']} nonlat={tot['nonlat']} "
          f"emptyStr={tot['empty_str']} err={tot['errores']}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--only', default=None, help='Solo esta carpeta/producto de dl/')
    ap.add_argument('--json', default=None, help='Vuelca la lista completa aquí')
    ap.add_argument('--fail', action='store_true', help='Exit 1 si hay defectos en algún producto')
    ap.add_argument('--quiet', action='store_true', help='No imprime la tabla, solo el resumen')
    args = ap.parse_args()

    items = recorrer(args.only)
    if not items:
        print(f'0 ficheros (only={args.only!r}) — ¿nombre de producto correcto?', file=sys.stderr)
        sys.exit(2 if args.fail else 0)

    rows = []
    for i, (prod, fname, path) in enumerate(items, 1):
        rows.append(censar_uno(prod, fname, path))
        if not args.quiet and i % 100 == 0:
            print(f'  ... {i}/{len(items)}', file=sys.stderr)

    prods = tabla_por_producto(rows)
    if not args.quiet:
        imprimir_tabla(prods)

    n_xlsx = sum(1 for r in rows if r.get('tipo') == 'xlsx')
    n_docx = sum(1 for r in rows if r.get('tipo') == 'docx')
    n_pdf = sum(1 for r in rows if r.get('tipo') == 'pdf')
    print(f'{len(rows)} ficheros censados ({n_xlsx} xlsx, {n_docx} docx, {n_pdf} pdf) '
          f'en {len(prods)} productos' + (f" (only={args.only})" if args.only else ''))

    if args.json:
        with open(args.json, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False, separators=(',', ':'))
        print(f'JSON escrito: {args.json} ({len(rows)} entradas)')

    if args.fail:
        defectuosos = [r for r in rows if r['prod'] != 'dl' and r.get('tipo') == 'xlsx' and es_defectuoso_xlsx(r)]
        if defectuosos:
            print(f'\nFAIL: {len(defectuosos)} xlsx con defectos:', file=sys.stderr)
            for r in defectuosos[:50]:
                razones = [k for k in DEFECTOS_XLSX if r.get(k, 0) > 0]
                if r.get('creator') != 'AI Chef Pro':
                    razones.append(f"creator={r.get('creator')!r}")
                if r.get('bio_vieja'):
                    razones.append('bio_vieja')
                if 'error' in r:
                    razones.append(f"error={r['error']}")
                print(f"  {r['prod']}/{r['f']}: {', '.join(razones)}", file=sys.stderr)
            if len(defectuosos) > 50:
                print(f'  ... y {len(defectuosos) - 50} más', file=sys.stderr)
            sys.exit(1)
        print('\nOK: 0 defectos (--fail)')
    sys.exit(0)


if __name__ == '__main__':
    main()
