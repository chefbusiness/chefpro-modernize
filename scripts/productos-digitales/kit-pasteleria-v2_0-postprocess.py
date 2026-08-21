#!/usr/bin/env python3
"""
Kit de Tareas Recurrentes · Pastelería / Obrador — post-proceso v2.0 (2026-08-21).

Trabaja SOBRE los .xlsx que ya están en astro-site/public/dl/kit-tareas-pasteleria/
(no regenera nada desde los generadores de marzo) y es IDEMPOTENTE: se puede
ejecutar N veces y el resultado es el mismo. Toda escritura es ABSOLUTA (nunca
"añade una fila más"), y los dos únicos desplazamientos que existen —la columna
nueva del 02 y la fila de 0,02 € del 09— van protegidos por un centinela que
detecta si ya se aplicaron.

Qué hace:
  1. 02-partidas-cocina    → columna «Parámetro / Objetivo» en las 3 hojas de
     partida + correcciones técnicas de la auditoría (fermentación, ganache,
     temperado, vapor, abatimiento, «glaseado espejo»).
  2. 06-eventos-festivos   → reconstruye las 4 campañas con las tareas que
     faltaban y añade «Todos los Santos (Oct-Nov)» y «Comuniones (Abr-Jun)».
     Cada campaña cierra con su fila de post-campaña.
  3. 09-apertura-cierre-caja → «Z del TPV» + «Descuadre» en el Registro Mensual,
     moneda de 0,02 € en el arqueo, formatos €, celdas de entrada en verde,
     validación de cantidades y formato condicional en DESCUADRE.
  4. Autorreferencias entre ficheros en 01-07 y BONUS-01: cada tarea que manda
     usar una ficha, un registro o un control cita el fichero por su número.
  5. Finalizador de los 15 ficheros: metadata v2.0, línea de versión, bio
     anclada, impresión A4 donde falte, limpieza de celdas vacías mal escritas
     y, al final del todo, cache de valores (inject_cache.py) + verificación.

Segunda pasada (auditoría adversarial de 3 lentes, 2026-08-22):
  6. 01 → horario de tienda coherente con el 08 (vitrinas a las 06:45 con
     comprobación de 2-6 °C, montaje 07:45, retirada de vitrina tras el cierre
     de las 19:50) y el producto del día anterior a promoción o a merma.
  7. 02 → cadena real del croissant: la 1ª hornada sale de la croissantería
     formada el día anterior (fermentación controlada nocturna) y la fila
     antigua pasa a ser la 2ª hornada de media mañana; fermentación 24-26 °C.
  8. 09 → el fondo de caja deja de contarse como venta: celda del fondo en la
     Apertura, fila «Fondo de caja inicial (−)» en el Resumen de Cierre y
     TOTAL FACTURADO = (efectivo − fondo) + tarjetas + otros; el fondo del
     Registro Mensual se promedia, no se suma; Instrucciones propias de caja.
  9. BONUS-02 → columna verde «Fecha de este año» y campaña de comuniones
     alineada con la hoja del 06.
 10. Los 26 checklists comparten vocabulario: encabezado «✓ Completada»,
     desplegable «✓ / — / N/A», relleno verde de la fila al marcar, contador
     con denominador calculado y 4 filas de holgura dentro del COUNTIF.

Nota para gates de ortografía: los dobles espacios de estos ficheros son
deliberados (sangría de las cabeceras de sección y separación de campos en las
líneas de firma con «___»); no son un defecto y no deben corregirse.

Uso:
    python3 scripts/productos-digitales/kit-pasteleria-v2_0-postprocess.py
    python3 scripts/productos-digitales/kit-pasteleria-v2_0-postprocess.py --skip 08,10,11,12,13
"""
import argparse
import copy
import os
import re
import subprocess
import sys

import datetime

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl', 'kit-tareas-pasteleria')
INJECT = os.path.join(ROOT, 'scripts', 'productos-digitales', 'inject_cache.py')

VERSION = '2.0'
VERSION_LINE = ('Versión 2.0 · agosto 2026 · aichef.pro/kit-tareas-pasteleria · '
                'info@aichef.pro')
RX_VERSION = re.compile(r'^Versión \d+\.\d+ · .*kit-tareas-pasteleria')
BIO_NEW = ('Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, '
           'en cocina desde los 17 años · johnguerrero.es')
PIE_HOJA = '— Kit de Tareas Recurrentes · Pastelería / Obrador · AI Chef Pro · aichef.pro'
SUBTITULO_TURNO = ('Fecha: ___/___/______    Turno: ☐ Mañana  ☐ Tarde  ☐ Noche    '
                   'Responsable turno: _________________________')

VERDE = 'E8F5E9'   # celdas editables
VERDE_OK = 'C8E6C9'  # fila de tarea completada (formato condicional)
AMBAR = 'FFF3CD'   # aviso de descuadre
FMT_EUR = '#,##0.00 €'

# Vocabulario único de los checklists del kit (SPEC §1.4 · L1-02)
MARCA_COL = '✓ Completada'
MARCA_COL_VIEJA = 'Hecha'
DV_LISTA = '"✓,—,N/A"'
HOLGURA = 4        # filas vacías, con desplegable y dentro del COUNTIF
LINEA_MARCA = ('Marca con ✓ en la columna «{}» (desplegable): es la que cuenta '
               'el total de tareas completadas.')

ETIQ_FONDO = 'Fondo de caja inicial (€)'
ETIQ_FONDO_RESUMEN = 'Fondo de caja inicial (−)'

NOMBRES = {
    '01': 'Apertura y Cierre',
    '02': 'Partidas de Producción',
    '03': 'Tareas del Manager',
    '04': 'Tareas por Perfil',
    '05': 'Tareas Semanales y Mensuales',
    '06': 'Eventos y Festivos',
    '07': 'Plantilla Personalizable',
    '08': 'Apertura y Cierre del Negocio',
    '09': 'Apertura y Cierre de Caja',
    '10': 'Plan de Producción Semanal + Mermas',
    '11': 'Control de Encargos',
    '12': 'Alérgenos de Vitrina (14 UE)',
    '13': 'Registro de Temperaturas, Recepción y Etiquetas',
}

COL_PARAM = 'Parámetro / Objetivo'

# --------------------------------------------------------------------------
# 4) Autorreferencias: marcador que se AÑADE al final del texto de la tarea.
# --------------------------------------------------------------------------
REF = {
    'ENCARGO': ' → 11 Control de Encargos · Ficha de Encargo',
    'TEMP': ' → 13 Registro de Temperaturas',
    'ETIQ': ' → 13 Registro de Temperaturas · Etiquetas de Elaborado',
    # OBR-10: el 10 no tiene ninguna pestaña «Mermas»; sus hojas son
    # Plan Semanal, Producido vs Vendido y Resumen por Partida.
    'MERMA': ' → 10 Plan de Producción Semanal · hoja Producido vs Vendido',
    'ALERG': ' → 12 Control de Alérgenos de Vitrina',  # nombre corto = título real del fichero 12 (R2-03)
    'CAJA': ' → 09 Apertura y Cierre de Caja',
}

AUTORREF = {
    '01-apertura-cierre.xlsx': [
        ('Encender las vitrinas refrigeradas y comprobar que están a 2-6 °C '
         'antes de montar', 'TEMP'),
        ('Verificar temperatura de cámaras frigoríficas (registrar)', 'TEMP'),
        ('Verificar temperaturas de cámaras (registrar en hoja de control)', 'TEMP'),
        ('Guardar masas en cámara de fermentación (etiquetar con fecha y hora)', 'ETIQ'),
        ('Guardar pre-elaboraciones en cámara (film, etiquetar, FIFO)', 'ETIQ'),
        ('Registrar mermas del día (productos descartados)', 'MERMA'),
        ('Preparar etiquetas de precio y alérgenos', 'ALERG'),
        ('Cuadrar caja / cierre de TPV', 'CAJA'),
    ],
    '02-partidas-cocina.xlsx': [
        ('Montar tartas por encargo (según la ficha de encargo de cada pedido)', 'ENCARGO'),
        ('Etiquetar todas las cremas con fecha y hora de elaboración', 'ETIQ'),
        ('Colocar etiquetas de alérgenos e ingredientes', 'ALERG'),
    ],
    '03-tareas-manager.xlsx': [
        ('Revisar pedidos especiales del día (tartas encargo, catering)', 'ENCARGO'),
        ('Gestionar pedidos de clientes (tartas personalizadas, eventos)', 'ENCARGO'),
        ('Revisar pedidos especiales de la semana (tartas, eventos)', 'ENCARGO'),
        ('Pedidos especiales pendientes (tartas encargo, catering)', 'ENCARGO'),
        ('Control de temperaturas de cámaras y vitrinas (registrar)', 'TEMP'),
        ('Revisar registros de temperatura de cámaras (últimos 3 días)', 'TEMP'),
        ('Revisar mermas del día (registrar producto descartado)', 'MERMA'),
        ('Analizar mermas del mes (objetivo: <5%)', 'MERMA'),
        ('Verificar etiquetado de alérgenos actualizado', 'ALERG'),
    ],
    '04-tareas-perfiles.xlsx': [
        ('Preparar pedidos especiales (tartas encargo)', 'ENCARGO'),
        ('Gestionar pedidos de tartas personalizadas (rellenar ficha)', 'ENCARGO'),
        ('Guardar productos en cámara (etiquetar)', 'ETIQ'),
        ('Verificar etiquetas de alérgenos y precios', 'ALERG'),
        ('Informar sobre ingredientes y alérgenos cuando pregunten', 'ALERG'),
        ('Cierre de caja / cuadre de TPV', 'CAJA'),
    ],
    '05-tareas-semanales-mensuales.xlsx': [
        ('Verificar temperaturas reales vs. display (sonda independiente)', 'TEMP'),
    ],
    '06-eventos-festivos.xlsx': [
        ('Control de pedidos de clientes (lista de encargos)', 'ENCARGO'),
        ('Preparar tartas de encargo personalizadas', 'ENCARGO'),
        ('Gestionar pedidos de encargo (monas, huevos personalizados)', 'ENCARGO'),
        ('Cerrar los encargos de Nochebuena y Navidad y repartir las franjas '
         'de recogida', 'ENCARGO'),
        ('Cerrar los encargos de roscón y repartir las franjas de recogida', 'ENCARGO'),
        ('Cerrar los encargos de mona de Pascua', 'ENCARGO'),
        ('Cerrar los encargos de bandejas y surtidos', 'ENCARGO'),
        ('Cerrar la ficha de encargo: sabores por piso, raciones, alérgenos, '
         'dedicatoria y hora de entrega', 'ENCARGO'),
        ('Cobrar la señal (30-50 %) y anotar el importe pendiente', 'ENCARGO'),
        ('Firmar la entrega con la familia y cobrar el pendiente', 'ENCARGO'),
        ('Etiquetar la vitrina de temporada: almendra, piñones y huevo están en '
         'casi todas las piezas', 'ALERG'),
        ('Preparar la mesa dulce: piezas, soportes, cartelería y etiquetas de '
         'alérgenos', 'ALERG'),
        ('Post-campaña: contar el sobrante, anotar las mermas y registrar qué se '
         'vendió y qué no', 'MERMA'),
    ],
    'BONUS-01-briefing-servicio.xlsx': [
        ('Tartas de encargo: nombre cliente, hora recogida, detalles', 'ENCARGO'),
        ('Alérgenos: productos con alérgenos especiales hoy', 'ALERG'),
    ],
}

# --------------------------------------------------------------------------
# 1) 02 — correcciones técnicas de texto (1:1 en la misma celda)
# --------------------------------------------------------------------------
FIXES_02 = {
    'Reposo en bloque en cámara (mínimo 1h a 4 °C)':
        'Bloque en frío antes de laminar (abatidor o cámara)',
    'Fermentar en cámara controlada (28 °C, 75% humedad, 1.5-2h)':
        'Fermentar en cámara controlada sin superar los 28 °C (por encima funde '
        'la mantequilla del laminado)',
    'Hornear con vapor (230-250 °C, vapor primeros 10 min)':
        'Hornear con vapor: golpe de vapor al cargar y abrir el tiro al final',
    'Enfriar rápidamente en baño maría inverso (abatir a <4 °C en 2h)':
        'Enfriar rápido en baño maría invertido o abatidor, tapar y etiquetar',
    'Preparar ganache de chocolate (55% para relleno, 70% para cobertura)':
        'Preparar ganache de relleno o de baño según el uso',
    'Temperar chocolate de cobertura (curva: 45°→27°→31° negro / 29° leche)':
        'Temperar el chocolate de cobertura siguiendo su curva',
    'Preparar glasé espejo (gelatina + glucosa + chocolate blanco + colorante)':
        'Preparar el glaseado espejo para el día siguiente (gelatina, glucosa, '
        'chocolate blanco y colorante)',
    # OBR-22: el glaseado necesita 12 h de reposo, así que se prepara D-1
    'Preparar glaseado espejo (gelatina, glucosa, chocolate blanco y colorante)':
        'Preparar el glaseado espejo para el día siguiente (gelatina, glucosa, '
        'chocolate blanco y colorante)',
    'Glasear con glasé espejo a 35 °C (sobre rejilla)':
        'Glasear con glaseado espejo sobre rejilla',
    # OBR-07: la fila antigua pasa a ser la 2ª hornada de media mañana
    'Cortar y dar forma (croissant, pain au chocolat, trenza)':
        'Cortar y dar forma: una parte a la 2ª hornada de hoy y otra al '
        'retardador para mañana',
    'Pincelar con huevo antes de hornear':
        'Pincelar con huevo la 2ª hornada antes de hornear',
    'Hornear (190-200 °C, 14-18 min según tamaño)':
        'Hornear la 2ª hornada de media mañana (190-200 °C, 14-18 min según tamaño)',
}

# 1) 02 — horas de la cadena del croissant (clave = texto YA corregido)
HORAS_02 = {
    'Sacar del retardador la croissantería formada ayer y comprobar el punto '
    'de fermentación': '06:45',
    'Pincelar y hornear la 1ª hornada del día para la apertura de la vitrina': '07:00',
    'Laminar con mantequilla (3 pliegues simples o 2 dobles)': '07:00',
    'Cortar y dar forma: una parte a la 2ª hornada de hoy y otra al retardador '
    'para mañana': '07:30',
    'Fermentar en cámara controlada sin superar los 28 °C (por encima funde la '
    'mantequilla del laminado)': '07:45',
    'Pincelar con huevo la 2ª hornada antes de hornear': '10:30',
    'Hornear la 2ª hornada de media mañana (190-200 °C, 14-18 min según tamaño)':
        '10:45',
}

# 1) 02 — las dos tareas nuevas de la 1ª hornada (se insertan antes de laminar)
CROISSANT_D1 = [
    ('Sacar del retardador la croissantería formada ayer y comprobar el punto '
     'de fermentación', 'Obrador', 'Pastelero'),
    ('Pincelar y hornear la 1ª hornada del día para la apertura de la vitrina',
     'Obrador', 'Pastelero'),
]
ANCLA_CROISSANT = 'Laminar con mantequilla (3 pliegues simples o 2 dobles)'

FIXES_09 = {
    '📋 Instrucciones de Uso': 'Instrucciones de Uso',
    '💡 Consejo: Plastifica una copia impresa y usa rotuladores borrables para '
    'reutilizarla cada día.':
        'Consejo: plastifica una copia impresa y usa rotuladores borrables para '
        'reutilizarla cada día.',
    '💰 Recuento de Efectivo por Denominación': 'Recuento de Efectivo por Denominación',
    '📊 Resumen de Cierre': 'Resumen de Cierre',
    # el enunciado de la tarea tiene que decir la fórmula que hay debajo
    'Registrar importe de fondo de caja en hoja':
        f'Anotar el importe del fondo en la celda «{ETIQ_FONDO}» de esta hoja',
    'Total facturado = efectivo + tarjetas + otros':
        'Total facturado = (efectivo contado − fondo de caja) + tarjetas + otros',
}

FIXES_06 = {
    '▸ Navidad, Reyes, San Valentín, Día de la Madre... cada una requiere planificación':
        '▸ Navidad, Reyes, San Valentín, Semana Santa, Día de la Madre, Comuniones '
        'y Todos los Santos: cada campaña tiene su hoja',
}

# --------------------------------------------------------------------------
# 1) 02 — parámetros por tarea (clave = texto YA corregido de la tarea)
# --------------------------------------------------------------------------
PARAMETROS_02 = {
    # Masas y Fermentación
    'Verificar estado de masa madre / prefermento (olor, volumen, textura)':
        'Dobla volumen en 3-4 h · olor láctico, nunca avinagrado',
    'Refrescar masa madre si es necesario (alimentar con harina y agua)':
        'Refresco 1:1:1 · reposo a 24-26 °C',
    'Pesar ingredientes para masa de croissant (harina, mantequilla, levadura)':
        'Harina W 300-330 · mantequilla de laminado 82 % MG',
    'Amasar corto en planetaria (1ª velocidad 3-4 min, 2ª velocidad 2-3 min): el '
    'gluten se termina de desarrollar en el laminado':
        '1ª vel. 3-4 min · 2ª vel. 2-3 min',
    'Control de temperatura de masa al salir (22-24 °C; 20-22 °C si va directa a '
    'bloque de frío)':
        '22-24 °C · 20-22 °C si va a bloque de frío',
    'Bloque en frío antes de laminar (abatidor o cámara)':
        '1 h en abatidor o 12 h en cámara a 3-4 °C',
    'Laminar con mantequilla (3 pliegues simples o 2 dobles)':
        'Masa y mantequilla a 12-14 °C · espesor final 3-3,5 mm',
    'Sacar del retardador la croissantería formada ayer y comprobar el punto '
    'de fermentación':
        'Fermentación controlada nocturna · sube a 24-26 °C · dobla volumen',
    'Pincelar y hornear la 1ª hornada del día para la apertura de la vitrina':
        '190-200 °C · 14-18 min · en la vitrina antes de las 08:00',
    'Cortar y dar forma: una parte a la 2ª hornada de hoy y otra al retardador '
    'para mañana':
        'Croissant 65-70 g · pain au chocolat 70-75 g',
    'Fermentar en cámara controlada sin superar los 28 °C (por encima funde la '
    'mantequilla del laminado)':
        '24-26 °C (28 °C techo) · 75-80 % HR · 2,5-3 h · dobla volumen',
    'Pincelar con huevo la 2ª hornada antes de hornear':
        'Huevo entero con una pizca de sal · capa fina',
    'Hornear la 2ª hornada de media mañana (190-200 °C, 14-18 min según tamaño)':
        '190-200 °C · 14-18 min',
    'Verificar fermentación nocturna de masas (volumen, alveolos)':
        '12-16 h a 4-6 °C',
    'Dividir y pesar porciones según formato (barra, hogaza, chapata)':
        'Barra 250-280 g · hogaza 800-900 g',
    'Segunda fermentación en banastones / telas (45 min - 1h)':
        '45-60 min a 24-26 °C',
    'Greñar / cortar con cuchilla antes de hornear':
        'Cuchilla inclinada 30-45°',
    'Hornear con vapor: golpe de vapor al cargar y abrir el tiro al final':
        '230-250 °C · vapor 3-5 min · tiro abierto los últimos 5-8 min',
    'Enfriar en rejillas (no apilar, mínimo 30 min)':
        'Mín. 30 min · sin apilar',
    'Preparar masa sablée (mantequilla pomada + azúcar + harina)':
        'Mantequilla pomada a 18-20 °C',
    'Preparar masa brisée para quiches/tartas saladas':
        'Reposo mín. 1 h a 4 °C',
    'Estirar y forrar moldes de tarta':
        'Espesor 2,5-3 mm',
    'Hornear en blanco con pesos (180 °C, 15 min)':
        '180 °C · 15 min',
    'Retirar pesos y dorar 5 min más':
        '180 °C · 5 min más',
    'Enfriar bases antes de rellenar':
        'Hasta 20-22 °C',
    # Cremas y Rellenos
    'Preparar crema pastelera (leche, yemas, azúcar, maicena, vainilla)':
        'Hervir 1-2 min sin dejar de remover',
    'Enfriar rápido en baño maría invertido o abatidor, tapar y etiquetar':
        'De +65 °C a +10 °C en menos de 2 h · después ≤4 °C',
    'Preparar crema de mantequilla (merengue suizo o italiana)':
        'Almíbar a 118-121 °C · mantequilla a 18-20 °C',
    'Preparar ganache de relleno o de baño según el uso':
        'Relleno 2:1 chocolate:nata · baño 1:1 · nata a 80-85 °C',
    'Preparar crema diplomática (pastelera + nata montada)':
        '2 partes de pastelera por 1 de nata montada',
    'Preparar lemon curd / curd de frutas de temporada':
        'Cocer hasta 82-84 °C',
    'Etiquetar todas las cremas con fecha y hora de elaboración':
        'Vida útil orientativa 48-72 h a ≤4 °C',
    'Preparar mousse de chocolate (templar chocolate + merengue + nata)':
        'Chocolate a 45-50 °C · mezcla final a 28-30 °C',
    'Preparar mousse de frutas (puré + gelatina + merengue + nata)':
        'Gelatina 10-12 g por litro · puré a 25-30 °C',
    'Preparar compota / confit de frutas para rellenos':
        'Cocer hasta 60-62 °Brix',
    'Preparar pralinés y praliné de frutos secos':
        'Tueste a 160 °C 12-15 min · caramelo a 170 °C',
    'Montar entremet en aros (capas de mousse + inserto + bizcocho)':
        'Aro de 16-18 cm · inserto congelado',
    'Congelar entremets montados para el glaseado de mañana (mín. 4 h en abatidor '
    'a −18 °C o 12 h en congelador)':
        'Mín. 4 h a −18 °C en abatidor · 12 h en congelador',
    'Preparar el glaseado espejo para el día siguiente (gelatina, glucosa, '
    'chocolate blanco y colorante)':
        'Reposo mínimo 12 h en frío antes de usarlo',
    'Temperar el chocolate de cobertura siguiendo su curva':
        'Negro 50-55 → 28-29 → 31-32 °C · leche 45 → 27-28 → 29-30 °C · '
        'blanco 40-45 → 26-27 → 28-29 °C',
    'Preparar fondant para glaseado de bollería':
        'Uso a 35-37 °C',
    'Preparar merengue italiano para decoración (almíbar 121 °C + claras)':
        'Almíbar a 121 °C',
    'Preparar caramelo para decoraciones (160-170 °C)':
        '160-170 °C',
    # Decoración y Acabado
    'Desmoldar los entremets congelados el día anterior':
        'Pieza a −18 °C · sin condensación',
    'Glasear con glaseado espejo sobre rejilla':
        'Glaseado a 32-35 °C · pieza a −18 °C',
    'Montar tartas por encargo (según la ficha de encargo de cada pedido)':
        'Raciones, dedicatoria y hora de recogida',
    'Montar vitrina de bollería (croissants, pain au chocolat, napolitanas)':
        'Ambiente · reponer en tandas pequeñas',
    'Montar vitrina de pastelería (tartas individuales, entremets, macarons)':
        'Vitrina refrigerada a 2-6 °C',
    'Montar vitrina de pan (si aplica)':
        'Ambiente seco · nunca en frío',
    'Colocar etiquetas de alérgenos e ingredientes':
        'Los 14 alérgenos de declaración obligatoria',
    'Reponer vitrina cada 2 horas durante servicio':
        'Cada 2 h · retirar lo que lleve más de un día',
}

# --------------------------------------------------------------------------
# 2) 06 — contenido completo de las campañas (se reconstruye entero)
#    Cada tarea = (Tarea, Zona, Responsable, Hora Límite)
# --------------------------------------------------------------------------
POST_CAMPANA = ('Post-campaña: contar el sobrante, anotar las mermas y registrar '
                'qué se vendió y qué no')

HOJAS_06 = [
    {
        'hoja': 'Navidad',
        'titulo': 'Temporada Navidad (Nov-Ene)',
        'tab': 'C62828',
        'secciones': [
            ('Planificación (Noviembre)', [
                ('Definir catálogo de Navidad (turrones, polvorones, roscón, troncos)',
                 'Admin', 'Jefe Pastelero', '1 Nov'),
                ('Calcular cantidades de producción por producto', 'Admin', 'Manager', '1 Nov'),
                ('Hacer pedidos especiales de ingredientes (mazapán, frutas confitadas, turrón)',
                 'Admin', 'Manager', '5 Nov'),
                ('Contratar personal extra si es necesario', 'Admin', 'Manager', '10 Nov'),
                ('Diseñar packaging navideño (cajas, lazos, etiquetas)', 'Admin', 'Manager', '10 Nov'),
                ('Publicar catálogo de Navidad en RRSS y web', 'Admin', 'Manager', '15 Nov'),
                ('Fijar y comunicar el horario del 24, 25 y 31 Dic y del 1 y 6 Ene',
                 'Admin', 'Manager', '20 Nov'),
            ]),
            ('Producción Navideña (Diciembre)', [
                ('Producción de turrones artesanos (2-3 semanas antes)', 'Obrador', 'Pastelero', '1-15 Dic'),
                ('Producción de polvorones y mantecados', 'Obrador', 'Pastelero', '1-15 Dic'),
                ('Producción de troncos de Navidad (bûche de Noël)', 'Obrador', 'Pastelero', '20-24 Dic'),
                ('Preparar cestas y lotes navideños', 'Obrador', 'Ayudante', '15-24 Dic'),
                ('Control de pedidos de clientes (lista de encargos)', 'Obrador', 'Dependiente', 'Continuo'),
            ]),
            ('Recogidas y Nochevieja (22-31 Dic)', [
                ('Cerrar los encargos de Nochebuena y Navidad y repartir las '
                 'franjas de recogida', 'Admin', 'Manager', '20 Dic'),
                ('Organizar la recogida de encargos por franjas horarias (22-24 Dic)',
                 'Vitrina', 'Dependiente', '22-24 Dic'),
                ('Reforzar el despacho en los picos del 23 y 24 (turno doble)',
                 'Vitrina', 'Manager', '23-24 Dic'),
                ('Cerrar los encargos de Nochevieja (postres y bandejas de dulces)',
                 'Admin', 'Manager', '29 Dic'),
                ('Producir y entregar los encargos del 31 de diciembre',
                 'Obrador', 'Pastelero', '31 Dic'),
            ]),
            ('Reyes (2-6 Ene)', [
                ('Cerrar los encargos de roscón y repartir las franjas de recogida',
                 'Admin', 'Manager', '2 Ene'),
                ('Preparar masas de roscón de Reyes (fermentación 24-48h)',
                 'Obrador', 'Pastelero', '3 Ene'),
                ('Hornear la 1ª hornada de roscones la madrugada del 4 al 5 (60-70 % '
                 'del volumen, para la venta de la tarde del 5)', 'Obrador', 'Pastelero', '04-05 Ene'),
                ('Hornear la 2ª hornada la madrugada del 5 al 6 y rellenar al momento',
                 'Obrador', 'Pastelero', '05-06 Ene'),
                ('Rellenar con nata, trufa, crema, cabello de ángel', 'Obrador', 'Pastelero', '05-06 Ene'),
                ('Decorar con fruta confitada, azúcar, corona', 'Obrador', 'Pastelero', '05-06 Ene'),
                ('Introducir sorpresa y haba', 'Obrador', 'Ayudante', '05-06 Ene'),
                ('Control de colas y turnos de recogida (reforzar despacho)',
                 'Vitrina', 'Dependiente', '05-06 Ene'),
            ]),
            ('Cierre de Campaña', [
                (POST_CAMPANA, 'Admin', 'Manager', '7 Ene'),
            ]),
        ],
    },
    {
        'hoja': 'San Valentín',
        'titulo': 'San Valentín (14 Febrero)',
        'tab': 'E91E63',
        'secciones': [
            ('Planificación (2 semanas antes)', [
                ('Diseñar colección San Valentín (tartas corazón, bombones, macarons)',
                 'Admin', 'Jefe Pastelero', '1 Feb'),
                ('Pedir moldes de corazón y packaging especial', 'Admin', 'Manager', '1 Feb'),
                ('Publicar en RRSS y abrir pedidos anticipados', 'Admin', 'Manager', '5 Feb'),
                ('Calcular producción extra (50-100% más que día normal)', 'Admin', 'Manager', '10 Feb'),
            ]),
            ('Producción (12-14 Feb)', [
                ('Producción de bombones y trufas artesanas', 'Obrador', 'Pastelero', '12 Feb'),
                ('Producción de tartas individuales con forma de corazón', 'Obrador', 'Pastelero', '13 Feb'),
                ('Producción de macarons con sabores especiales (rosa, frambuesa, pasión)',
                 'Obrador', 'Pastelero', '13 Feb'),
                ('Montar cajas regalo con selección de piezas', 'Obrador', 'Ayudante', '13 Feb'),
                ('Preparar tartas de encargo personalizadas', 'Obrador', 'Pastelero', '14 Feb'),
                ('Reforzar despacho para pico de ventas (14 Feb tarde)', 'Obrador', 'Dependiente', '14 Feb'),
            ]),
            ('Cierre de Campaña', [
                (POST_CAMPANA, 'Admin', 'Manager', '15 Feb'),
            ]),
        ],
    },
    {
        'hoja': 'Semana Santa',
        'titulo': 'Semana Santa (Cuaresma y Pascua)',
        'tab': 'FF6F00',
        'secciones': [
            ('Planificación', [
                ('Definir producción de Semana Santa (torrijas, monas, huevos de Pascua)',
                 'Admin', 'Jefe Pastelero', '4 sem antes'),
                ('Pedir moldes de huevos de Pascua y colorantes', 'Admin', 'Manager', '3 sem antes'),
                ('Producción de figuras de chocolate (conejos, huevos, campanas)',
                 'Admin', 'Pastelero', '2 sem antes'),
                ('Cerrar los encargos de mona de Pascua', 'Admin', 'Manager', 'Dom Ramos'),
            ]),
            ('Producción', [
                ('Elaborar torrijas a diario (remojo, fritura, almíbar/canela) — pico '
                 'Jueves y Viernes Santo', 'Obrador', 'Pastelero', 'V. Dolores → D. Resurrección'),
                ('Elaborar pestiños y leche frita (fritura, almíbar y canela)',
                 'Obrador', 'Pastelero', 'Mié-Vie Santo'),
                ('Freír buñuelos de Cuaresma y rellenarlos al momento',
                 'Obrador', 'Pastelero', 'Vie Santo'),
                ('Elaborar y montar monas de Pascua (bizcocho + decoración chocolate) '
                 'para recogida el Domingo de Resurrección y el Lunes de Pascua',
                 'Obrador', 'Pastelero', 'Jue-Sáb Santo'),
                ('Producir huevos de chocolate rellenos', 'Obrador', 'Pastelero', 'Jue-Sáb Santo'),
                ('Hornear hornazo para el Lunes de Pascua (según la zona)',
                 'Obrador', 'Pastelero', 'Sáb Santo'),
            ]),
            ('Venta y Encargos', [
                ('Montar escaparate temático de Semana Santa', 'Vitrina', 'Dependiente', 'Dom Ramos'),
                ('Gestionar pedidos de encargo (monas, huevos personalizados)',
                 'Vitrina', 'Dependiente', 'Continuo'),
            ]),
            ('Cierre de Campaña', [
                (POST_CAMPANA, 'Admin', 'Manager', 'Mar tras Pascua'),
            ]),
        ],
    },
    {
        'hoja': 'Día Madre-Padre',
        'titulo': 'Día del Padre (19 de marzo) y Día de la Madre (primer domingo de mayo — '
                  'fecha de este año: ____/____/________)',
        'tab': '9C27B0',
        'secciones': [
            ('Preparación', [
                ('Anotar la fecha de este año del Día de la Madre (primer domingo de mayo)',
                 'Admin', 'Manager', '3 sem antes'),
                ('Diseñar tartas especiales para la ocasión', 'Admin', 'Jefe Pastelero', '2 sem antes'),
                ('Abrir pedidos de tartas personalizadas (mensaje, dedicatoria)',
                 'Admin', 'Manager', '2 sem antes'),
                ('Preparar packaging especial', 'Admin', 'Manager', '1 sem antes'),
            ]),
            ('Producción y Venta', [
                ('Producción extra de tartas (incremento 80-120%)', 'Obrador', 'Pastelero', 'Día antes'),
                ('Preparar tartas con dedicatoria personalizada', 'Obrador', 'Pastelero', 'Día'),
                ('Reforzar despacho (turno doble si es necesario)', 'Obrador', 'Dependiente', 'Día'),
                ('Ofrecer servicio de entrega a domicilio (si aplica)', 'Obrador', 'Manager', 'Día'),
            ]),
            ('Cierre de Campaña', [
                (POST_CAMPANA, 'Admin', 'Manager', 'Día siguiente'),
            ]),
        ],
    },
    {
        'hoja': 'Comuniones (Abr-Jun)',
        'titulo': 'Comuniones (Abril-Junio)',
        'tab': '1976D2',
        'secciones': [
            ('Captación y Visita (Enero-Marzo)', [
                ('Preparar el catálogo de comuniones (tartas por pisos, mesas dulces, '
                 'detalles) con fotos y precios', 'Admin', 'Jefe Pastelero', 'Enero'),
                ('Publicar el catálogo y abrir la agenda de visitas', 'Admin', 'Manager', 'Febrero'),
                ('Visita con la familia: fecha, nº de invitados, lugar, estilo y '
                 'presupuesto orientativo', 'Admin', 'Manager', 'Feb-Mar'),
            ]),
            ('Presupuesto, Prueba y Señal', [
                ('Enviar el presupuesto por escrito detallando qué incluye y qué no '
                 '(transporte, montaje, alquiler de soportes)', 'Admin', 'Manager', '48 h tras la visita'),
                ('Prueba de sabores con la familia (2-3 combinaciones)',
                 'Obrador', 'Jefe Pastelero', '4-6 sem antes'),
                ('Cerrar la ficha de encargo: sabores por piso, raciones, alérgenos, '
                 'dedicatoria y hora de entrega', 'Admin', 'Manager', '4 sem antes'),
                ('Cobrar la señal (30-50 %) y anotar el importe pendiente',
                 'Admin', 'Manager', 'Al confirmar'),
                ('Confirmar con la familia el nº definitivo de invitados y la hora',
                 'Admin', 'Manager', '7 días antes'),
            ]),
            ('Producción y Montaje', [
                ('Elaborar bizcochos y rellenos, y calar las plantillas de cada piso',
                 'Obrador', 'Pastelero', '2 días antes'),
                ('Montar los pisos con estructura interna y revisar el aplomo de la pieza',
                 'Obrador', 'Jefe Pastelero', 'Día antes'),
                ('Decorar y guardar la tarta montada en cámara sin olores',
                 'Obrador', 'Pastelero', 'Día antes'),
                ('Preparar la mesa dulce: piezas, soportes, cartelería y etiquetas de '
                 'alérgenos', 'Obrador', 'Ayudante', 'Día antes'),
            ]),
            ('Transporte y Entrega', [
                ('Preparar el transporte: caja rígida, base antideslizante, vehículo '
                 'refrigerado y ruta con margen', 'Obrador', 'Manager', 'Día de la entrega'),
                ('Salir del obrador con 60 min de margen sobre la hora acordada',
                 'Obrador', 'Pastelero', 'Día de la entrega'),
                ('Montaje in situ, retoque final y foto de la pieza terminada',
                 'Obrador', 'Jefe Pastelero', 'Día de la entrega'),
                ('Firmar la entrega con la familia y cobrar el pendiente',
                 'Admin', 'Manager', 'Día de la entrega'),
            ]),
            ('Cierre de Campaña', [
                (POST_CAMPANA, 'Admin', 'Manager', 'Julio'),
            ]),
        ],
    },
    {
        'hoja': 'Todos los Santos (Oct-Nov)',
        'titulo': 'Todos los Santos (Octubre-Noviembre)',
        'tab': '6D4C41',
        'secciones': [
            ('Planificación (Octubre)', [
                ('Definir la colección de temporada (huesos de santo, panellets, '
                 'buñuelos de viento)', 'Admin', 'Jefe Pastelero', '10 Oct'),
                ('Pedir mazapán, almendra molida, piñones y boniato con antelación',
                 'Admin', 'Manager', '12 Oct'),
                ('Publicar la colección en RRSS y abrir los encargos', 'Admin', 'Manager', '20 Oct'),
                ('Calcular la producción por día del 28 Oct al 2 Nov', 'Admin', 'Manager', '20 Oct'),
            ]),
            ('Producción (28 Oct - 2 Nov)', [
                ('Elaborar la masa de mazapán y formar los huesos de santo (yema, '
                 'cabello de ángel, chocolate)', 'Obrador', 'Pastelero', '28-31 Oct'),
                ('Elaborar panellets (piñón, almendra, coco, café)', 'Obrador', 'Pastelero', '28-31 Oct'),
                ('Freír los buñuelos de viento y rellenarlos al momento (crema, nata, '
                 'chocolate)', 'Obrador', 'Pastelero', '31 Oct-1 Nov'),
                ('Etiquetar la vitrina de temporada: almendra, piñones y huevo están en '
                 'casi todas las piezas', 'Vitrina', 'Dependiente', '31 Oct'),
                ('Montar el escaparate de temporada y el cartel de precios por peso',
                 'Vitrina', 'Dependiente', '30 Oct'),
            ]),
            ('Encargos y Venta (30 Oct - 2 Nov)', [
                ('Cerrar los encargos de bandejas y surtidos', 'Admin', 'Manager', '30 Oct'),
                ('Organizar las recogidas del 1 de noviembre por franjas horarias',
                 'Vitrina', 'Dependiente', '31 Oct'),
                ('Reforzar el despacho el 1 de noviembre (turno doble)', 'Vitrina', 'Manager', '1 Nov'),
                ('Reponer la vitrina cada 2 h: el buñuelo se vende recién relleno',
                 'Vitrina', 'Dependiente', '1-2 Nov'),
            ]),
            ('Cierre de Campaña', [
                (POST_CAMPANA, 'Admin', 'Manager', '3 Nov'),
            ]),
        ],
    },
]

RX_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')


# ==========================================================================
# Utilidades genéricas
# ==========================================================================
def _traducir_formula(valor, idx, eje):
    """Desplaza en 1 las referencias de columna (eje='col') o de fila (eje='fila')
    iguales o posteriores a idx dentro de una fórmula."""
    if not (isinstance(valor, str) and valor.startswith('=')):
        return valor

    def _sub(m):
        d1, col, d2, fila = m.groups()
        ci = column_index_from_string(col)
        fi = int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        return f'{d1}{col}{d2}{fila}'

    return RX_REF.sub(_sub, valor)


# Atributos que hay que conservar al reconstruir una validación: si solo se
# copian type/formula1/allow_blank se pierden en silencio los mensajes de error
# (el «Cantidad no válida» del arqueo del 09).
CAMPOS_DV = ('type', 'formula1', 'formula2', 'operator', 'allow_blank',
             'showErrorMessage', 'errorTitle', 'error', 'errorStyle',
             'showInputMessage', 'promptTitle', 'prompt', 'showDropDown')


def _rangos_dv(ws):
    return [({k: getattr(dv, k, None) for k in CAMPOS_DV},
             [str(r) for r in dv.sqref.ranges])
            for dv in ws.data_validations.dataValidation]


def _restaurar_dv(ws, guardados, idx=None, eje=None):
    ws.data_validations.dataValidation = []
    for attrs, rangos in guardados:
        dv = DataValidation(**{k: v for k, v in attrs.items() if v is not None})
        ws.add_data_validation(dv)
        for r in rangos:
            dv.add(_desplazar_rango(r, idx, eje) if idx else r)


def _desplazar_rango(ref, idx, eje):
    partes = ref.split(':')
    fuera = []
    for p in partes:
        m = RX_REF.fullmatch(p)
        if not m:
            return ref
        d1, col, d2, fila = m.groups()
        ci, fi = column_index_from_string(col), int(fila)
        if eje == 'col' and ci >= idx:
            col = get_column_letter(ci + 1)
        if eje == 'fila' and fi >= idx:
            fila = str(fi + 1)
        fuera.append(f'{d1}{col}{d2}{fila}')
    return ':'.join(fuera)


def insertar_columna(ws, idx):
    """Inserta una columna en idx manteniendo a mano lo que openpyxl NO mueve:
    combinaciones, validaciones, fórmulas y anchos de columna."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    anchos = {k: v.width for k, v in ws.column_dimensions.items() if v.width}

    for col in range(max_c, idx - 1, -1):
        for fila in range(1, max_r + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila, column=col + 1)
            dst.value = _traducir_formula(src.value, idx, 'col')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'col'))
    _restaurar_dv(ws, dvs, idx, 'col')

    for letra, ancho in sorted(anchos.items(), key=lambda kv: -column_index_from_string(kv[0])):
        ci = column_index_from_string(letra)
        if ci >= idx:
            ws.column_dimensions[get_column_letter(ci + 1)].width = ancho


def insertar_fila(ws, idx):
    """Equivalente por filas de insertar_columna."""
    max_c, max_r = ws.max_column, ws.max_row
    merges = [str(r) for r in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    dvs = _rangos_dv(ws)
    alturas = {k: v.height for k, v in ws.row_dimensions.items() if v.height}

    for fila in range(max_r, idx - 1, -1):
        for col in range(1, max_c + 1):
            src = ws.cell(row=fila, column=col)
            dst = ws.cell(row=fila + 1, column=col)
            dst.value = _traducir_formula(src.value, idx, 'fila')
            dst._style = copy.copy(src._style)
            src.value = None

    for m in merges:
        ws.merge_cells(_desplazar_rango(m, idx, 'fila'))
    _restaurar_dv(ws, dvs, idx, 'fila')

    for fila, alto in sorted(alturas.items(), reverse=True):
        if fila >= idx:
            ws.row_dimensions[fila + 1].height = alto


def print_setup(ws, header_row=None, landscape=True):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    if header_row:
        ws.print_title_rows = f'{header_row}:{header_row}'
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def fila_cabecera(ws):
    for r in range(1, 8):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if b == 'Tarea' or a in ('#', 'Nº', 'Fecha', 'Denominación'):
            return r
    return None


def base_texto(v):
    """Texto de la tarea sin el marcador de autorreferencia."""
    return v.split(' → ')[0] if isinstance(v, str) else v


def _es_fila_seccion(ws, fila):
    """True si la fila es una cabecera de sección (celda combinada A:…)."""
    for m in ws.merged_cells.ranges:
        if m.min_row == fila and m.max_row == fila and m.min_col == 1 and m.max_col >= 3:
            return True
    return False


def _buscar(ws, texto, col=None):
    """Fila donde aparece exactamente `texto` (en `col` o en cualquier columna)."""
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if c.value == texto:
                return c.row
    return None


def _buscar_prefijo(ws, prefijo, col=1):
    """Fila cuya celda empieza por `prefijo`. Se usa donde el texto lleva
    rellenos de guiones bajos y espacios múltiples (líneas de firma), que son
    deliberados y no conviene reproducir a mano en el código."""
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if isinstance(c.value, str) and c.value.startswith(prefijo):
                return c.row
    return None


def actualizar_tarea(ws, claves, texto=None, hora=None, col_tarea=2, col_hora=5):
    """Localiza una tarea por su texto base (admite varias grafías, para ser
    idempotente) y le cambia el enunciado y/o la hora límite."""
    for r in range(1, ws.max_row + 1):
        cel = ws.cell(row=r, column=col_tarea)
        if not isinstance(cel.value, str) or base_texto(cel.value) not in claves:
            continue
        if texto is not None and cel.value != texto:
            cel.value = texto
        if hora is not None:
            ws.cell(row=r, column=col_hora).value = hora
        return r
    return None


def renumerar_tareas(ws, filas_extra=()):
    """Renumera la columna Nº de una hoja de checklist (1..n) contando las filas
    de tarea existentes más las recién insertadas."""
    n = 0
    for r in range(1, ws.max_row + 1):
        cel = ws.cell(row=r, column=1)
        if isinstance(cel.value, int) or r in filas_extra:
            n += 1
            cel.value = n
    return n


def _limpiar_cf(ws, marca):
    """Elimina las reglas de formato condicional cuya fórmula contiene `marca`
    (para poder reescribirlas sin duplicarlas ni tocar las demás)."""
    nuevas = {}
    for k, reglas in ws.conditional_formatting._cf_rules.items():
        quedan = [r for r in reglas
                  if not any(marca in str(f) for f in (getattr(r, 'formula', None) or []))]
        if quedan:
            nuevas[k] = quedan
    ws.conditional_formatting._cf_rules = nuevas


# ==========================================================================
# Vocabulario único de los checklists (T-06/T-07/T-12 · OBR-19 · COM-05)
# ==========================================================================
def analizar_checklist(ws):
    """Geometría de una hoja de checklist, o None si no lo es."""
    hr = None
    for r in range(1, 8):
        if ws.cell(row=r, column=2).value == 'Tarea':
            hr = r
            break
    if hr is None:
        return None
    col_marca = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hr, column=c).value
        if isinstance(v, str) and v.strip() in (MARCA_COL_VIEJA, MARCA_COL):
            col_marca = c
    if col_marca is None:
        return None
    contador = _buscar(ws, 'Tareas completadas:', col=1)
    ultima = None
    for r in range(hr + 1, (contador or ws.max_row + 1)):
        if isinstance(ws.cell(row=r, column=1).value, int):
            ultima = r
    if ultima is None:
        return None
    return {'hr': hr, 'col_marca': col_marca, 'ultima': ultima, 'contador': contador}


def normalizar_checklist(ws):
    """Encabezado «✓ Completada», desplegable ✓/—/N/A, relleno verde de la fila
    al marcar, 4 filas de holgura y contador con numerador Y denominador
    calculados. Idempotente: si la hoja ya está normalizada no cambia nada."""
    g = analizar_checklist(ws)
    if not g:
        return None
    hr, cm, ultima, contador = g['hr'], g['col_marca'], g['ultima'], g['contador']

    ws.cell(row=hr, column=cm).value = MARCA_COL

    # --- filas de holgura: 4 vacías entre la última tarea y el contador ----
    if contador:
        hueco = contador - 1 - ultima
        for _ in range(max(0, (HOLGURA + 1) - hueco)):
            insertar_fila(ws, ultima + 1)
            contador += 1
        fin = ultima + HOLGURA
        for r in range(ultima + 1, fin + 1):
            for c in range(1, ws.max_column + 1):
                cel = ws.cell(row=r, column=c)
                cel.value = None
                cel._style = copy.copy(ws.cell(row=ultima, column=c)._style)
    else:
        fin = ultima

    # --- validación de datos ---------------------------------------------
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if dv.type != 'list']
    dv = DataValidation(type='list', formula1=DV_LISTA, allow_blank=True)
    ws.add_data_validation(dv)
    for r in range(hr + 1, fin + 1):
        if isinstance(ws.cell(row=r, column=1).value, int) or ultima < r <= fin:
            dv.add(ws.cell(row=r, column=cm))

    # --- formato condicional: fila verde cuando la marca es ✓ -------------
    letra = get_column_letter(cm)
    _limpiar_cf(ws, '"✓"')
    rango = f'A{hr + 1}:{get_column_letter(ws.max_column)}{fin}'
    ws.conditional_formatting.add(rango, FormulaRule(
        formula=[f'${letra}{hr + 1}="✓"'],
        fill=PatternFill('solid', start_color=VERDE_OK, end_color=VERDE_OK)))

    # --- contador ---------------------------------------------------------
    if contador:
        col_num = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=contador, column=c).value == 'de':
                col_num = c - 1   # numerador | «de» | denominador
                break
        if col_num is None:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=contador, column=c).value
                if isinstance(v, str) and v.startswith('=COUNTIF'):
                    col_num = c
                    break
        if col_num:
            ws.cell(row=contador, column=col_num).value = (
                f'=COUNTIF({letra}{hr + 1}:{letra}{fin},"✓")')
            ws.cell(row=contador, column=col_num + 1).value = 'de'
            # denominador calculado sobre la columna Tarea: si el usuario
            # añade o borra tareas, el total deja de mentir (T-07)
            ws.cell(row=contador, column=col_num + 2).value = (
                f'=COUNTIF(B{hr + 1}:B{fin},"?*")')
    return ws.title


def normalizar_checklists(wb, informe):
    hechas = [t for t in (normalizar_checklist(ws) for ws in wb.worksheets) if t]
    if hechas:
        informe.append(f'    checklists normalizados ({len(hechas)}): '
                       + ', '.join(f'«{h}»' for h in hechas))
    return hechas


def unificar_texto_instrucciones(wb):
    """La línea de Instrucciones nombra la columna por su nombre real."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and f'columna «{MARCA_COL_VIEJA}»' in c.value:
                    c.value = c.value.replace(f'columna «{MARCA_COL_VIEJA}»',
                                              f'columna «{MARCA_COL}»')


# ==========================================================================
# 1) 02 — columna «Parámetro / Objetivo»
# ==========================================================================
def _cadena_croissant(ws, informe):
    """OBR-07: la vitrina abre a las 08:00, así que la 1ª hornada sale de la
    croissantería formada el día anterior. Inserta las dos tareas que faltaban
    antes del laminado y ajusta las horas de la 2ª hornada."""
    if _buscar(ws, CROISSANT_D1[0][0], col=2) is None:
        ancla = _buscar(ws, ANCLA_CROISSANT, col=2)
        if ancla is None:
            print(f'    AVISO: no encuentro «{ANCLA_CROISSANT[:40]}» en «{ws.title}»')
            return
        nuevas = []
        for i, (tarea, zona, resp) in enumerate(CROISSANT_D1):
            fila = ancla + i
            insertar_fila(ws, fila)
            ws.cell(row=fila, column=2).value = tarea
            ws.cell(row=fila, column=4).value = zona
            ws.cell(row=fila, column=5).value = resp
            nuevas.append(fila)
        renumerar_tareas(ws, filas_extra=set(nuevas))
        informe.append(f'    «{ws.title}»: 1ª hornada del día (2 tareas) '
                       'antes del laminado')
    for r in range(5, ws.max_row + 1):
        clave = base_texto(ws.cell(row=r, column=2).value)
        if clave in HORAS_02:
            ws.cell(row=r, column=6).value = HORAS_02[clave]


def ajustar_02(wb, informe):
    hojas = ['Masas y Fermentación', 'Cremas y Rellenos', 'Decoración y Acabado']
    for titulo in hojas:
        ws = wb[titulo]
        if ws.cell(row=4, column=3).value != COL_PARAM:
            insertar_columna(ws, 3)
            ws.cell(row=4, column=3).value = COL_PARAM
            ws.cell(row=4, column=3)._style = copy.copy(ws.cell(row=4, column=2)._style)
            for r in range(5, ws.max_row + 1):
                if isinstance(ws.cell(row=r, column=1).value, int):
                    ws.cell(row=r, column=3)._style = copy.copy(ws.cell(row=r, column=2)._style)
            ws.column_dimensions['C'].width = 34
            informe.append(f'    columna «{COL_PARAM}» insertada en «{titulo}» (C)')
        if titulo == 'Masas y Fermentación':
            _cadena_croissant(ws, informe)
        # relleno de parámetros (idempotente)
        puestos = 0
        for r in range(5, ws.max_row + 1):
            if not isinstance(ws.cell(row=r, column=1).value, int):
                continue
            clave = base_texto(ws.cell(row=r, column=2).value)
            par = PARAMETROS_02.get(clave)
            if par:
                ws.cell(row=r, column=3).value = par
                puestos += 1
        informe.append(f'    «{titulo}»: {puestos} parámetros rellenados')
        print_setup(ws, 4, landscape=True)


# ==========================================================================
# 2) 06 — reconstrucción de las campañas
# ==========================================================================
def _plantilla_06(ws):
    """Estilos de referencia de una campaña. Las filas se localizan por su
    CONTENIDO, no por su número: la hoja se reconstruye en cada pasada y con
    números fijos la segunda ejecución copiaría los estilos de otras filas."""
    secciones, dato = [], None
    for r in range(5, ws.max_row + 1):
        if _es_fila_seccion(ws, r):
            secciones.append(r)
        elif dato is None and isinstance(ws.cell(row=r, column=1).value, int):
            dato = r
    contador = _buscar(ws, 'Tareas completadas:', col=1) or 29
    verif = _buscar(ws, 'Verificado por:', col=1) or (contador + 2)
    pie = _buscar(ws, PIE_HOJA, col=1) or (verif + 2)
    sec_a = secciones[0] if secciones else 5
    sec_b = secciones[1] if len(secciones) > 1 else sec_a
    return {
        'titulo': copy.copy(ws.cell(row=1, column=1)._style),
        'subtitulo': copy.copy(ws.cell(row=2, column=1)._style),
        'cab': [copy.copy(ws.cell(row=4, column=c)._style) for c in range(1, 8)],
        'sec_a': [copy.copy(ws.cell(row=sec_a, column=c)._style) for c in range(1, 8)],
        'sec_b': [copy.copy(ws.cell(row=sec_b, column=c)._style) for c in range(1, 8)],
        'dato': [copy.copy(ws.cell(row=dato or 6, column=c)._style) for c in range(1, 8)],
        'contador': [copy.copy(ws.cell(row=contador, column=c)._style) for c in range(1, 8)],
        'verif': [copy.copy(ws.cell(row=verif, column=c)._style) for c in range(1, 8)],
        'pie': copy.copy(ws.cell(row=pie, column=1)._style),
        'anchos': {k: v.width for k, v in ws.column_dimensions.items() if v.width},
    }


def _escribir_campana(ws, spec, P):
    CAB = ['Nº', 'Tarea', 'Zona', 'Responsable', 'Hora Límite', MARCA_COL, 'Firma']
    ws.cell(row=1, column=1, value=spec['titulo'])._style = copy.copy(P['titulo'])
    ws.merge_cells('A1:G1')
    ws.cell(row=2, column=1, value=SUBTITULO_TURNO)._style = copy.copy(P['subtitulo'])
    ws.merge_cells('A2:G2')
    for c, texto in enumerate(CAB, start=1):
        cel = ws.cell(row=4, column=c, value=texto)
        cel._style = copy.copy(P['cab'][c - 1])

    dv = DataValidation(type='list', formula1=DV_LISTA, allow_blank=True)
    ws.add_data_validation(dv)

    fila = 5
    n = 0
    for i, (seccion, tareas) in enumerate(spec['secciones']):
        estilos = P['sec_a'] if i == 0 else P['sec_b']
        for c in range(1, 8):
            ws.cell(row=fila, column=c)._style = copy.copy(estilos[c - 1])
        ws.cell(row=fila, column=1).value = f'  {seccion}'
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
        fila += 1
        for tarea, zona, resp, hora in tareas:
            n += 1
            for c in range(1, 8):
                ws.cell(row=fila, column=c)._style = copy.copy(P['dato'][c - 1])
            ws.cell(row=fila, column=1).value = n
            ws.cell(row=fila, column=2).value = tarea
            ws.cell(row=fila, column=3).value = zona
            ws.cell(row=fila, column=4).value = resp
            ws.cell(row=fila, column=5).value = hora
            dv.add(ws.cell(row=fila, column=6))
            fila += 1
        fila += 1  # línea en blanco entre secciones

    ultima = fila - 2          # última fila con tarea
    holgura = ultima + 2       # filas de holgura dentro del COUNTIF
    contador = ultima + 3
    for c in range(1, 8):
        ws.cell(row=contador, column=c)._style = copy.copy(P['contador'][c - 1])
    ws.cell(row=contador, column=1).value = 'Tareas completadas:'
    ws.merge_cells(start_row=contador, start_column=1, end_row=contador, end_column=3)
    ws.cell(row=contador, column=4).value = f'=COUNTIF(F5:F{holgura},"✓")'
    ws.cell(row=contador, column=5).value = 'de'
    ws.cell(row=contador, column=6).value = n

    verif = contador + 2
    for c in range(1, 8):
        ws.cell(row=verif, column=c)._style = copy.copy(P['verif'][c - 1])
    ws.cell(row=verif, column=1).value = 'Verificado por:'
    ws.merge_cells(start_row=verif, start_column=2, end_row=verif, end_column=4)
    ws.cell(row=verif, column=5).value = 'Firma:'
    ws.merge_cells(start_row=verif, start_column=6, end_row=verif, end_column=7)

    pie = verif + 2
    ws.cell(row=pie, column=1, value=PIE_HOJA)._style = copy.copy(P['pie'])

    for letra, ancho in P['anchos'].items():
        ws.column_dimensions[letra].width = ancho
    ws.sheet_properties.tabColor = spec['tab']
    print_setup(ws, 4, landscape=True)
    return n


def ajustar_06(wb, informe):
    P = _plantilla_06(wb['Navidad'])
    for i, spec in enumerate(HOJAS_06, start=1):
        nombre = spec['hoja']
        nuevo = nombre not in wb.sheetnames
        if not nuevo:
            del wb[nombre]
        ws = wb.create_sheet(nombre, i)
        n = _escribir_campana(ws, spec, P)
        informe.append(f'    «{nombre}»: {n} tareas'
                       + (' (hoja NUEVA)' if nuevo else ' (reconstruida)'))


# ==========================================================================
# 2 bis) 01 — horario de tienda coherente con el 08 (OBR-05/06/09)
# ==========================================================================
TXT_VITRINAS = ('Encender las vitrinas refrigeradas y comprobar que están a '
                '2-6 °C antes de montar')
TXT_RETIRAR_VITRINA = ('Retirar el producto de vitrina y separar el sobrante: a promoción '
                       'señalizada o a merma, nunca a la vitrina normal de mañana'
                       ' → 10 Plan de Producción Semanal · hoja Producido vs Vendido')
TXT_DIA_ANTERIOR = ('Colocar la producción del día; el producto del día anterior '
                    'solo en la zona de promoción señalizada o a merma'
                    ' → 10 Plan de Producción Semanal · hoja Producido vs Vendido'
                    ' · 13 Registro de Temperaturas · Vidas Útiles')


def ajustar_01(wb, informe):
    ap = wb['Apertura Obrador']
    # Las vitrinas necesitan bajar a régimen ANTES de montar el género: se
    # encienden a las 06:45 y se comprueba la temperatura (el 08 usa la misma
    # regla). El montaje es a las 07:45 y la tienda abre a las 08:00.
    actualizar_tarea(ap, {'Encender iluminación y climatización de vitrinas',
                          TXT_VITRINAS},
                     texto=TXT_VITRINAS, hora='06:45')
    actualizar_tarea(ap, {'Colocar productos del día anterior (verificar caducidad)',
                          base_texto(TXT_DIA_ANTERIOR)},
                     texto=TXT_DIA_ANTERIOR, hora='07:45')

    ci = wb['Cierre Obrador']
    # La vitrina se retira DESPUÉS de cerrar al público (19:50) y la caja se
    # cuadra a las 20:10, la misma hora que el 08 (R2-02). El sobrante NO vuelve a
    # la vitrina normal de mañana: promoción señalizada o merma, la misma regla que
    # ya fija la apertura (R2-01).
    actualizar_tarea(ci, {'Retirar productos de vitrina (valorar si se pueden '
                          'vender mañana)', base_texto(TXT_RETIRAR_VITRINA)},
                     texto=TXT_RETIRAR_VITRINA, hora='19:55')
    actualizar_tarea(ci, {'Cuadrar caja / cierre de TPV'}, hora='20:10')
    informe.append('    horario de tienda alineado con el 08 (vitrinas 06:45, '
                   'montaje 07:45, retirada 19:55, caja 20:10)')


# ==========================================================================
# 2 ter) BONUS-01 — el briefing se usa dos veces al día (OBR-21)
# ==========================================================================
BULLET_BONUS01 = ('▸ Se usa dos veces al día: al arrancar el obrador y en la '
                  'apertura de la tienda → 08 Apertura y Cierre del Negocio')


def ajustar_bonus01(wb, informe):
    ws = wb['Briefing Diario']
    n = 0
    for r in range(5, ws.max_row + 1):
        if isinstance(ws.cell(row=r, column=1).value, int):
            if ws.cell(row=r, column=5).value != 'Inicio de turno':
                ws.cell(row=r, column=5).value = 'Inicio de turno'
                n += 1
    ins = wb['Instrucciones']
    if _buscar(ins, BULLET_BONUS01) is None:
        ancla = _buscar(ins, '▸ Imprime una copia cada día y pégala en el obrador')
        if ancla:
            insertar_fila(ins, ancla + 1)
            cel = ins.cell(row=ancla + 1, column=2, value=BULLET_BONUS01)
            cel._style = copy.copy(ins.cell(row=ancla, column=2)._style)
    if n:
        informe.append(f'    Briefing Diario: {n} horas fijas → «Inicio de turno» '
                       '(el briefing se hace en los dos arranques)')


# ==========================================================================
# 2 quater) BONUS-02 — fecha del año en curso y campaña de comuniones
# ==========================================================================
COL_FECHA_ANIO = 'Fecha de este año'
FECHAS_BONUS02 = {          # texto del evento → (Fecha, Antelación, Notas)
    'Semana Santa': ('Mar-Abr (variable)', None, None),
    'Día de la Madre': ('1er dom. de mayo (variable)', None, None),
    'Comuniones': ('Abr-Jun (variable)', '4-5 meses (catálogo en enero)',
                   'Pedidos con mucha antelación → 06 Eventos y Festivos · '
                   'hoja Comuniones (Abr-Jun)'),
}


PREFIJO_BONUS02 = '▸ Anota en la columna verde «Fecha de este año»'
BULLET_BONUS02 = (PREFIJO_BONUS02 + ' el día en que caen Semana Santa, el Día '
                  'de la Madre, las comuniones y el Black Friday')


def ajustar_bonus02(wb, informe):
    ins = wb['Instrucciones']
    fila = _buscar_prefijo(ins, PREFIJO_BONUS02, col=2)
    if fila is None:
        ancla = _buscar(ins, '▸ Cada fecha incluye qué producir y con cuánta antelación')
        if ancla:
            insertar_fila(ins, ancla + 1)
            fila = ancla + 1
            ins.cell(row=fila, column=2)._style = copy.copy(
                ins.cell(row=ancla, column=2)._style)
    if fila:
        ins.cell(row=fila, column=2).value = BULLET_BONUS02

    ws = wb['Calendario Anual']
    cab = 3
    if ws.cell(row=cab, column=3).value != COL_FECHA_ANIO:
        insertar_columna(ws, 3)
        cel = ws.cell(row=cab, column=3, value=COL_FECHA_ANIO)
        cel._style = copy.copy(ws.cell(row=cab, column=2)._style)
        ws.column_dimensions['C'].width = 20
        informe.append(f'    columna «{COL_FECHA_ANIO}» insertada (C, editable)')
    for r in range(cab + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ''):
            continue
        if not isinstance(ws.cell(row=r, column=4).value, str):
            continue
        cel = ws.cell(row=r, column=3)
        cel._style = copy.copy(ws.cell(row=r, column=2)._style)
        cel.value = None
        _verde(cel)
    # los números de orden son números, no texto (T-16)
    n = 0
    for r in range(cab + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.isdigit():
            ws.cell(row=r, column=1).value = int(v)
            n += 1
    if n:
        informe.append(f'    columna Nº: {n} valores de texto convertidos a número')
    # fechas móviles marcadas y campaña de comuniones alineada con el 06
    for r in range(cab + 1, ws.max_row + 1):
        evento = ws.cell(row=r, column=4).value
        if evento in FECHAS_BONUS02:
            fecha, antelacion, notas = FECHAS_BONUS02[evento]
            ws.cell(row=r, column=2).value = fecha
            if antelacion:
                ws.cell(row=r, column=6).value = antelacion
            if notas:
                ws.cell(row=r, column=7).value = notas
    informe.append('    fechas móviles marcadas como «(variable)» y comuniones '
                   'remitidas a la hoja del 06')


# ==========================================================================
# 3) 09 — caja
# ==========================================================================
def _verde(cel):
    cel.fill = PatternFill('solid', fgColor=VERDE)


INSTRUCCIONES_09 = [
    (2, '09 — Apertura y Cierre de Caja · Pastelería / Obrador', 'titulo'),
    (4, 'Qué resuelve:', 'cabecera'),
    (5, '▸ El arqueo diario de la caja de la tienda: fondo inicial, recuento por '
        'denominación y cuadre contra la Z del TPV', 'texto'),
    (6, '▸ El registro mensual, para ver de un vistazo qué días descuadraron y '
        'cuánto', 'texto'),
    (8, 'Cómo usar:', 'cabecera'),
    (9, '▸ Apertura de Caja: cuenta el fondo y anótalo en la celda verde «Fondo '
        'de caja inicial (€)»', 'texto'),
    (10, '▸ Cierre de Caja: saca la Z del TPV, cuenta el efectivo por denominación '
         'y anota las tarjetas y los otros medios de pago', 'texto'),
    (11, '▸ El fondo se descuenta solo: total facturado = (efectivo contado − '
         'fondo) + tarjetas + otros', 'texto'),
    (12, '▸ Si la casilla DESCUADRE se pone ámbar, la caja no cuadra: revísala '
         'antes de cerrar', 'texto'),
    (13, '▸ Registro Mensual: anota cada día el efectivo, las tarjetas, los otros '
         'medios y la Z; el descuadre se calcula solo', 'texto'),
    (15, 'Personalización:', 'cabecera'),
    (16, '▸ Las celdas verdes son editables — ajusta importes, responsables y '
         'horarios', 'texto'),
    (17, '▸ Añade o elimina tareas según tu operativa y borra las denominaciones '
         'que no uses', 'texto'),
    (18, '▸ Plastifica una copia impresa y usa rotuladores borrables para '
         'reutilizarla cada día', 'texto'),
    (20, 'Se conecta con:', 'cabecera'),
    (21, '▸ 08 Apertura y Cierre del Negocio — el cierre de caja es una de sus '
         'últimas tareas del día', 'texto'),
    (22, '▸ 10 Plan de Producción Semanal — el sobrante de la vitrina se registra '
         'en su hoja Producido vs Vendido', 'texto'),
    (24, f'Marca con ✓ en la columna «{MARCA_COL}» (desplegable) cada tarea que '
         'completes.', 'texto'),
    (26, '© AI Chef Pro — aichef.pro', 'pie'),
    (27, BIO_NEW, 'pie'),
    (28, 'Contacto: info@aichef.pro', 'pie'),
    (30, VERSION_LINE, 'pie'),
]


def _estilo_de(ws, textos):
    """Estilo de la primera celda que contenga uno de esos textos. Se busca por
    CONTENIDO para que la función siga encontrando sus modelos después de haber
    reescrito la hoja una vez (los textos originales ya no están donde estaban)."""
    for t in textos:
        for row in ws.iter_rows():
            for c in row:
                if c.value == t:
                    return copy.copy(c._style)
    return None


def _instrucciones_09(wb, informe):
    """COM-12: las Instrucciones del 09 eran las genéricas del 08, en la columna
    A y sin una sola referencia al resto del kit. Se reescriben con el molde de
    los demás ficheros (columna B, Qué resuelve / Cómo usar / Personalización /
    Se conecta con)."""
    ws = wb['Instrucciones']
    esperado = {f: t for f, t, _ in INSTRUCCIONES_09}
    actual = {r: ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)
              if ws.cell(row=r, column=2).value is not None}
    if actual == esperado:
        return
    estilos = {
        'titulo': _estilo_de(ws, ['Checklist de Apertura y Cierre de Caja',
                                  INSTRUCCIONES_09[0][1]]),
        'cabecera': _estilo_de(ws, ['Instrucciones de Uso', 'Qué resuelve:']),
        'texto': _estilo_de(ws, [
            '1. Este archivo contiene checklists profesionales para la apertura '
            'y cierre de tu negocio.',
            '▸ El registro mensual, para ver de un vistazo qué días descuadraron '
            'y cuánto']),
        'pie': _estilo_de(ws, [BIO_NEW]),
    }
    for m in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(m)
    for r in range(1, ws.max_row + 1):
        for c in range(1, max(ws.max_column, 2) + 1):
            ws.cell(row=r, column=c).value = None
    for fila, texto, estilo in INSTRUCCIONES_09:
        cel = ws.cell(row=fila, column=2, value=texto)
        if estilos[estilo] is not None:
            cel._style = copy.copy(estilos[estilo])
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 108
    informe.append('    Instrucciones reescritas para la caja (molde del resto '
                   'del kit, con las referencias al 08 y al 10)')


def _fondo_de_caja(wb, informe):
    """OBR-01/T-10: el fondo de caja no es una venta. Se anota en la Apertura,
    se descuenta del recuento en el Resumen de Cierre y en el Registro Mensual
    se promedia en vez de sumarse (es un saldo, no un flujo)."""
    ap = wb['Apertura de Caja']
    fila_fondo = _buscar(ap, ETIQ_FONDO, col=2)
    if fila_fondo is None:
        firma = _buscar_prefijo(ap, 'Firma del responsable:', col=1)
        if firma is None:
            firma = ap.max_row - 1
        for _ in range(2):
            insertar_fila(ap, firma)
        fila_fondo = firma
        modelo = firma - 2                      # última fila de tarea
        blanca = firma - 1                      # fila en blanco que la separa
        for c in range(1, ap.max_column + 1):
            ap.cell(row=fila_fondo, column=c)._style = copy.copy(
                ap.cell(row=modelo, column=c)._style)
            ap.cell(row=fila_fondo + 1, column=c).value = None
            ap.cell(row=fila_fondo + 1, column=c)._style = copy.copy(
                ap.cell(row=blanca, column=c)._style)
        ap.cell(row=fila_fondo, column=2).value = ETIQ_FONDO
        cel = ap.cell(row=fila_fondo, column=3)
        cel.value = 0
        cel.number_format = FMT_EUR
        _verde(cel)
        ap.cell(row=fila_fondo, column=4).value = (
            'Se descuenta del efectivo en el Cierre de Caja')
        ap.merge_cells(start_row=fila_fondo, start_column=4,
                       end_row=fila_fondo, end_column=6)
        informe.append(f'    Apertura de Caja: celda del fondo de caja '
                       f'(C{fila_fondo}, editable)')

    ci = wb['Cierre de Caja']
    r_ef = _buscar(ci, 'Total Efectivo', col=1) or _buscar(
        ci, 'Total Efectivo (recuento)', col=1)
    if r_ef is None:
        print('    AVISO: no encuentro «Total Efectivo» en el Resumen de Cierre')
        return
    if ci.cell(row=r_ef + 1, column=1).value != ETIQ_FONDO_RESUMEN:
        insertar_fila(ci, r_ef + 1)
        informe.append('    Cierre de Caja: fila «Fondo de caja inicial (−)» en '
                       'el Resumen de Cierre')
    # el fondo es una celda CALCULADA (viene de la Apertura): se viste como
    # «Total Efectivo», no en verde, que en este kit significa «escribe aquí»
    for c in range(1, ci.max_column + 1):
        ci.cell(row=r_ef + 1, column=c)._style = copy.copy(
            ci.cell(row=r_ef, column=c)._style)
    ci.cell(row=r_ef, column=1).value = 'Total Efectivo (recuento)'
    ci.cell(row=r_ef + 1, column=1).value = ETIQ_FONDO_RESUMEN
    cel = ci.cell(row=r_ef + 1, column=2)
    cel.value = f"=IFERROR('Apertura de Caja'!C{fila_fondo},0)"
    cel.number_format = FMT_EUR

    r_tar = _buscar(ci, 'Total Tarjetas (Visa/MC)', col=1)
    r_otr = _buscar(ci, 'Total Otros (Bizum, Vales)', col=1)
    r_tot = _buscar(ci, 'TOTAL FACTURADO', col=1)
    r_z = _buscar(ci, 'Z del TPV', col=1)
    r_desc = _buscar(ci, 'DESCUADRE', col=1)
    if None in (r_tar, r_otr, r_tot, r_z, r_desc):
        print('    AVISO: falta alguna fila del Resumen de Cierre del 09')
        return
    ci.cell(row=r_tot, column=2).value = (
        f'=IFERROR(B{r_ef}-B{r_ef + 1},0)+B{r_tar}+B{r_otr}')
    ci.cell(row=r_desc, column=2).value = f'=IFERROR(B{r_tot}-B{r_z},0)'
    for r in (r_ef, r_ef + 1, r_tot, r_desc):
        ci.cell(row=r, column=2).number_format = FMT_EUR

    # --- Registro Mensual: el fondo es un saldo, no una venta -------------
    rm = wb['Registro Mensual']
    # AVERAGEIF y no AVERAGE: las 31 filas nacen a 0 y promediar los días que
    # todavía no se han rellenado devolvería una cifra sin sentido.
    rm.cell(row=36, column=2).value = '=IFERROR(AVERAGEIF(B5:B35,">0"),0)'
    rm.cell(row=36, column=2).number_format = FMT_EUR
    nota = ('El fondo de caja es un saldo, no una venta: en la fila TOTALES la '
            'columna «Fondo Apertura» muestra el fondo medio de los días '
            'anotados, no su suma.')
    if rm.cell(row=37, column=1).value != nota:
        cel = rm.cell(row=37, column=1, value=nota)
        cel._style = copy.copy(rm.cell(row=38, column=1)._style)
        if 'A37:J37' not in [str(r) for r in rm.merged_cells.ranges]:
            rm.merge_cells('A37:J37')
        informe.append('    Registro Mensual: el fondo pasa a promedio («fondo '
                       'medio») con su nota al pie')


def ajustar_09(wb, informe):
    # --- Apertura y Cierre: celdas de entrada en verde -------------------
    for hoja in ('Apertura de Caja', 'Cierre de Caja'):
        ws = wb[hoja]
        for r in range(5, ws.max_row + 1):
            if isinstance(ws.cell(row=r, column=1).value, int):
                for c in (3, 4, 5, 6):
                    _verde(ws.cell(row=r, column=c))

    _instrucciones_09(wb, informe)

    ws = wb['Cierre de Caja']
    # --- moneda de 0,02 € (auditoría L1-02/L1-23) ------------------------
    if ws.cell(row=35, column=1).value != '0,02 €':
        insertar_fila(ws, 35)
        for c in range(1, 4):
            ws.cell(row=35, column=c)._style = copy.copy(ws.cell(row=34, column=c)._style)
        ws.cell(row=35, column=1).value = '0,02 €'
        ws.cell(row=35, column=2).value = 0
        ws.cell(row=35, column=3).value = '=B35*0.02'
        ws.cell(row=35, column=3).number_format = FMT_EUR
        informe.append('    Cierre de Caja: añadida la denominación de 0,02 €')

    _fondo_de_caja(wb, informe)

    # --- cantidades: enteros ≥ 0, verdes ---------------------------------
    dv_cant = None
    for dv in ws.data_validations.dataValidation:
        if dv.type == 'whole':
            dv_cant = dv
    if dv_cant is None:
        dv_cant = DataValidation(type='whole', operator='greaterThanOrEqual',
                                 formula1='0', allow_blank=True,
                                 showErrorMessage=True,
                                 errorTitle='Cantidad no válida',
                                 error='Escribe el número de billetes o monedas '
                                       '(un número entero, sin decimales).')
        ws.add_data_validation(dv_cant)
    for r in range(21, 37):
        if isinstance(ws.cell(row=r, column=1).value, str) and '€' in ws.cell(row=r, column=1).value:
            cel = ws.cell(row=r, column=2)
            cel.number_format = '0'
            _verde(cel)
            dv_cant.add(cel)

    # --- resumen de cierre: verdes en lo editable + condicional ----------
    fila_desc = None
    for r in range(37, ws.max_row + 1):
        etiqueta = ws.cell(row=r, column=1).value
        if etiqueta in ('Total Tarjetas (Visa/MC)', 'Total Otros (Bizum, Vales)', 'Z del TPV'):
            cel = ws.cell(row=r, column=2)
            cel.number_format = FMT_EUR
            _verde(cel)
        if etiqueta == 'DESCUADRE':
            fila_desc = r
    if fila_desc:
        ref = f'B{fila_desc}'
        # se descartan TODAS las reglas de una sola celda de la columna B: si
        # la fila del DESCUADRE se ha movido, la vieja quedaría huérfana
        # pintando de ámbar una celda que ya no es la del descuadre.
        ws.conditional_formatting._cf_rules = {
            k: v for k, v in ws.conditional_formatting._cf_rules.items()
            if not re.fullmatch(r'B\d+', str(k.sqref))
        }
        ws.conditional_formatting.add(ref, CellIsRule(
            operator='notEqual', formula=['0'],
            fill=PatternFill('solid', start_color=AMBAR, end_color=AMBAR)))
        informe.append(f'    Cierre de Caja: formato condicional ámbar en DESCUADRE ({ref})')

    # --- Registro Mensual: Z del TPV + Descuadre -------------------------
    rm = wb['Registro Mensual']
    est_cab = copy.copy(rm.cell(row=4, column=6)._style)
    est_num = copy.copy(rm.cell(row=5, column=2)._style)
    est_form = copy.copy(rm.cell(row=5, column=6)._style)
    est_resp = copy.copy(rm.cell(row=5, column=9)._style)
    est_tot = copy.copy(rm.cell(row=36, column=6)._style)

    for col, texto in ((7, 'Z del TPV'), (8, 'Descuadre'), (9, 'Depósito'), (10, 'Responsable')):
        cel = rm.cell(row=4, column=col, value=texto)
        cel._style = copy.copy(est_cab)

    for r in range(5, 36):
        g = rm.cell(row=r, column=7)
        g._style = copy.copy(est_num)
        g.value = 0
        g.number_format = FMT_EUR
        _verde(g)

        h = rm.cell(row=r, column=8)
        h._style = copy.copy(est_form)
        h.value = f'=IFERROR(F{r}-G{r},0)'
        h.number_format = FMT_EUR

        i = rm.cell(row=r, column=9)
        i._style = copy.copy(est_num)
        i.value = 0
        i.number_format = FMT_EUR
        _verde(i)

        j = rm.cell(row=r, column=10)
        j._style = copy.copy(est_resp)
        j.value = None
        _verde(j)

        for c in range(2, 7):
            celda = rm.cell(row=r, column=c)
            celda.number_format = FMT_EUR
            if c != 6:
                _verde(celda)

    for col in (7, 8, 9):
        cel = rm.cell(row=36, column=col)
        cel._style = copy.copy(est_tot)
        letra = get_column_letter(col)
        cel.value = f'=SUM({letra}5:{letra}35)'
        cel.number_format = FMT_EUR
    cel = rm.cell(row=36, column=10)
    cel._style = copy.copy(est_tot)
    cel.value = None

    for ref_viejo, ref_nuevo in (('A1:I1', 'A1:J1'), ('A38:I38', 'A38:J38'), ('F2:I2', 'F2:J2')):
        if ref_viejo in [str(r) for r in rm.merged_cells.ranges]:
            rm.unmerge_cells(ref_viejo)
        if ref_nuevo not in [str(r) for r in rm.merged_cells.ranges]:
            rm.merge_cells(ref_nuevo)
    rm.column_dimensions['G'].width = 14
    rm.column_dimensions['H'].width = 14
    rm.column_dimensions['I'].width = 14
    rm.column_dimensions['J'].width = 16

    ws_desc = 'H5:H35'
    rm.conditional_formatting._cf_rules = {
        k: v for k, v in rm.conditional_formatting._cf_rules.items()
        if str(k.sqref) != ws_desc
    }
    rm.conditional_formatting.add(ws_desc, CellIsRule(
        operator='notEqual', formula=['0'],
        fill=PatternFill('solid', start_color=AMBAR, end_color=AMBAR)))
    informe.append('    Registro Mensual: columnas «Z del TPV» y «Descuadre» '
                   '(=IFERROR(Total−Z,0)) con formato € y aviso de descuadre')

    # --- alturas fijas + wrap (auditoría L1-15) --------------------------
    for hoja in wb.worksheets:
        for r, dim in list(hoja.row_dimensions.items()):
            if dim.height and dim.height < 30:
                if any(hoja.cell(row=r, column=c).alignment.wrap_text
                       for c in range(1, hoja.max_column + 1)):
                    dim.height = None


# ==========================================================================
# 4) Autorreferencias
# ==========================================================================
def aplicar_autorreferencias(wb, fname, inventario):
    reglas = AUTORREF.get(fname, [])
    if not reglas:
        return
    pendientes = {texto: clave for texto, clave in reglas}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str):
                    continue
                base = base_texto(c.value)
                clave = pendientes.get(base) or dict(reglas).get(base)
                if not clave:
                    continue
                marcador = REF[clave]
                if marcador in c.value:
                    inventario.append((fname, ws.title, c.coordinate, clave, 'ya presente'))
                else:
                    c.value = base + marcador
                    inventario.append((fname, ws.title, c.coordinate, clave, 'aplicada'))
                pendientes.pop(base, None)
    for texto in pendientes:
        print(f'    AVISO: autorreferencia sin destino en {fname}: «{texto[:70]}»')


# ==========================================================================
# 5) Finalizador común
# ==========================================================================
def set_metadata(wb, fname):
    num = fname.split('-')[0]
    if fname.startswith('BONUS-01'):
        titulo = 'Bonus 01 — Briefing de Servicio · Kit de Tareas Pastelería / Obrador'
    elif fname.startswith('BONUS-02'):
        titulo = 'Bonus 02 — Calendario Anual · Kit de Tareas Pastelería / Obrador'
    else:
        titulo = f'{num} — {NOMBRES.get(num, num)} · Kit de Tareas Pastelería / Obrador'
    p = wb.properties
    p.creator = 'AI Chef Pro'
    p.lastModifiedBy = 'AI Chef Pro'
    p.title = titulo
    p.subject = f'Kit de Tareas Recurrentes · Pastelería / Obrador · v{VERSION}'
    p.keywords = 'pastelería, obrador, checklist, tareas, AI Chef Pro'
    p.description = 'aichef.pro/kit-tareas-pasteleria'
    p.category = 'AI Chef Pro · Productos digitales'


def linea_instrucciones(ws, texto, rx=None):
    """Escribe `texto` en Instrucciones: sustituye la línea que case con `rx`
    (línea de versión) o la añade al final si no existe. Nunca duplica."""
    col = 2 if ws.cell(row=2, column=2).value else 1
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                return
    destino = ws.max_row + 2
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)


def finalizar(wb, fname):
    set_metadata(wb, fname)
    for ws in wb.worksheets:
        # celdas '' → None (XML no conforme, auditoría L1-14)
        for row in ws.iter_rows():
            for c in row:
                if c.value == '':
                    c.value = None
                elif isinstance(c.value, str) and ('29 años' in c.value or '15 años' in c.value):
                    c.value = BIO_NEW
        if ws.title == 'Instrucciones':
            linea_instrucciones(ws, VERSION_LINE, RX_VERSION)
            if ws.page_setup.paperSize is None:
                ws.page_setup.paperSize = 9
                ws.page_setup.orientation = 'portrait'
            continue
        if ws.page_setup.paperSize is None:
            hr = fila_cabecera(ws)
            print_setup(ws, hr, landscape=ws.max_column >= 6)


# ==========================================================================
# Proceso por fichero
# ==========================================================================
def procesar(fname, inventario):
    path = os.path.join(DL, fname)
    wb = openpyxl.load_workbook(path)
    informe = []

    fixes = {}
    if fname == '02-partidas-cocina.xlsx':
        fixes = FIXES_02
    elif fname == '06-eventos-festivos.xlsx':
        fixes = FIXES_06
    elif fname == '09-apertura-cierre-caja.xlsx':
        fixes = FIXES_09
    if fixes:
        n = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value in fixes:
                        c.value = fixes[c.value]
                        n += 1
        if n:
            informe.append(f'    {n} correcciones de texto')

    if fname == '01-apertura-cierre.xlsx':
        ajustar_01(wb, informe)
    elif fname == '02-partidas-cocina.xlsx':
        ajustar_02(wb, informe)
    elif fname == '06-eventos-festivos.xlsx':
        ajustar_06(wb, informe)
    elif fname == '09-apertura-cierre-caja.xlsx':
        ajustar_09(wb, informe)
    elif fname == 'BONUS-01-briefing-servicio.xlsx':
        ajustar_bonus01(wb, informe)
    elif fname == 'BONUS-02-calendario-anual-tareas.xlsx':
        ajustar_bonus02(wb, informe)

    aplicar_autorreferencias(wb, fname, inventario)
    normalizar_checklists(wb, informe)
    unificar_texto_instrucciones(wb)
    finalizar(wb, fname)
    wb.save(path)
    return informe


# ==========================================================================
# Verificación
# ==========================================================================
RX_NO_LATINO = re.compile('[぀-ヿ㐀-䶿一-鿿가-힯'
                          'Ѐ-ӿ؀-ۿ֐-׿฀-๿]')
PROHIBIDOS = ('29 años', '15 años', 'openpyxl', 'ChefBusiness')


def verificar(fname):
    path = os.path.join(DL, fname)
    wbv = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path)
    formulas = no_latinos = fechas = 0
    pendientes = []          # fórmulas sin valor cacheado, por clasificar
    prohibidos = []
    version_ok = False
    for ws, wsf in zip(wbv.worksheets, wbf.worksheets):
        for row in wsf.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
                    fechas += 1
                if isinstance(v, str):
                    if v.startswith('='):
                        formulas += 1
                        if ws[c.coordinate].value is None:
                            pendientes.append((wsf.title, c.coordinate))
                    if RX_NO_LATINO.search(v):
                        no_latinos += 1
                    for p in PROHIBIDOS:
                        if p in v:
                            prohibidos.append(f'{ws.title}!{c.coordinate}:{p}')
                    if v == VERSION_LINE:
                        version_ok = True

    # Una fórmula que evalúa a cadena vacía NO es un fallo de cache: es lo que
    # hacen los IF(...="","",…) de la carta de alérgenos, y contarlas en rojo
    # acabaría con el equipo ignorando el rojo (T-09). Solo se paga el coste de
    # pycel cuando queda alguna fórmula sin valor.
    sin_cache, vacias = 0, 0
    if pendientes:
        from pycel import ExcelCompiler
        xl = ExcelCompiler(filename=path)
        for hoja, ref in pendientes:
            try:
                valor = xl.evaluate(f"'{hoja}'!{ref}")
            except Exception:
                valor = None
            if isinstance(valor, str) and valor == '':
                vacias += 1
            else:
                sin_cache += 1

    return {
        'hojas': len(wbf.worksheets),
        'formulas': formulas,
        'sin_cache': sin_cache,
        'vacias_por_diseno': vacias,
        'fechas': fechas,
        'no_latinos': no_latinos,
        'prohibidos': prohibidos,
        'creator': wbf.properties.creator,
        'subject': wbf.properties.subject,
        'titulo': wbf.properties.title,
        'version_v2': version_ok,
    }


def main():
    ap = argparse.ArgumentParser(description='Post-proceso v2.0 del Kit de Tareas Pastelería')
    ap.add_argument('--skip', default='', help='números de fichero a excluir, p. ej. 08,10,11,12,13')
    args = ap.parse_args()
    saltar = {s.strip() for s in args.skip.split(',') if s.strip()}

    todos = sorted(f for f in os.listdir(DL) if f.endswith('.xlsx') and not f.startswith('~$'))
    ficheros = [f for f in todos if f.split('-')[0] not in saltar]
    omitidos = [f for f in todos if f not in ficheros]
    print(f'{len(ficheros)} ficheros a procesar en {DL}')
    if omitidos:
        print('  omitidos por --skip: ' + ', '.join(omitidos))

    inventario = []
    for f in ficheros:
        print(f'  {f}')
        for linea in procesar(f, inventario):
            print(linea)

    print('\nINVENTARIO DE AUTORREFERENCIAS')
    for fich, hoja, coord, clave, estado in inventario:
        print(f'  {fich}:{hoja}:{coord}  {REF[clave].strip()}  [{estado}]')
    print(f'  total: {len(inventario)}')

    print('\nCACHE DE VALORES (inject_cache.py)')
    subprocess.run([sys.executable, INJECT] + [os.path.join(DL, f) for f in ficheros], check=True)

    print('\nVERIFICACIÓN')
    ok = True
    for f in ficheros:
        v = verificar(f)
        bien = (v['sin_cache'] == 0 and v['no_latinos'] == 0 and v['fechas'] == 0
                and not v['prohibidos']
                and v['creator'] == 'AI Chef Pro' and v['subject'].endswith('v2.0')
                and v['version_v2'])
        ok &= bien
        print(f"  {'OK  ' if bien else 'FALLA'} {f}: {v}")
    print('\nTODO OK' if ok else '\nHAY FALLOS')
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
