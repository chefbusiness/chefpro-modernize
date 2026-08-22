#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
postprocess-transversal.py — Fase A del saneamiento de los entregables .xlsx
============================================================================

Aplica a los 42 productos de `astro-site/public/dl/` (todos menos
`kit-tareas-pasteleria`, que ya está en v2.0, y los ficheros sueltos de la raíz
de `/dl/`, que son huérfanos) los arreglos DETERMINISTAS del censo del
2026-08-22, sin tocar una sola cifra ni una sola tarea de negocio:

  §2.1 metadata de marca          §2.4 vocabulario único de la casilla
  §2.2 bio anclada                §2.5 línea de versión en Instrucciones
  §2.3 impresión A4               §2.6 limpieza, cache de valores y verificación

Especificación: `scripts/productos-digitales/FASE-A-SPEC-postprocess-transversal.md`
Patrón de referencia: `kit-pasteleria-v1_1-postprocess.py` y `…v2_0-postprocess.py`.

Uso
---
    python3 scripts/productos-digitales/postprocess-transversal.py \
        <productId>|all|<ruta/a/carpeta> [--dry-run] [--json informe.json] [--skip-cache]

  --dry-run    copia la carpeta del producto a `$SCRATCH/dryrun/<pid>/`, procesa
               ALLÍ y deja `/dl/` intacto. Es el modo con el que se valida.
  --json       vuelca el informe completo (cambios por fichero + verificación).
  --skip-cache no ejecuta `inject_cache.py` (útil para iterar rápido); en ese
               modo `nocache_real` y `contadores_ko` se informan pero NO tumban
               la verificación: sin cache son falsos positivos por construcción.
  --nombres-cortos  imprime el nombre corto de cada producto y termina; sirve
               para demostrar que ninguno colisiona con otro (§2.1).

Se puede pasar una RUTA de carpeta en vez de un productId: procesa esa carpeta
in place. Es lo que permite demostrar la idempotencia (segunda pasada sobre la
copia que dejó `--dry-run`, cuyo informe debe traer 0 cambios). Si la ruta cae
DENTRO de `/dl/`, el veto de `EXCLUIDOS` se aplica igual que con el productId.

Exit code: 0 si todas las verificaciones de §2.6 pasan, 1 si falla alguna.

Lo que este script NO hace (§3): cambiar contenido de negocio, insertar filas o
columnas, renombrar ficheros, tocar docx/pdf, tocar la marca «ChefBusiness» de
los productos de origen CB, ni tocar `kit-tareas-pasteleria`.

REGLA TÉRMICA: procesa en serie, un fichero detrás de otro. No paralelizar.
"""

import argparse
import collections
import copy
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ==========================================================================
# Constantes
# ==========================================================================
AQUI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(AQUI))
DL = os.path.join(ROOT, 'astro-site', 'public', 'dl')
PRODUCTOS_TS = os.path.join(ROOT, 'astro-site', 'src', 'data', 'productos')
INJECT_CACHE = os.path.join(AQUI, 'inject_cache.py')

SCRATCH = os.environ.get(
    'FASE_A_SCRATCH',
    '/private/tmp/claude-501/-Users-johnguerrero-chefpro-modernize/'
    'f83a04d3-d14f-49b3-aa79-c442fb4d7983/scratchpad')

#: Producto ya migrado a v2.0 con su propio post-proceso: intocable aquí.
EXCLUIDOS = {'kit-tareas-pasteleria'}

VERSION = '1.1'
MES = 'agosto 2026'

# --- §2.4 vocabulario único de la casilla (el de pastelería v2.0) ---------
MARCA_COL = '✓ Completada'
CABECERAS_MARCA = ('Hecha', '✓ Completada', '✓', 'OK')
DV_TAREA = '"✓,—,N/A"'      # columna de marca separada (P1/P2/P3b)
DV_CASILLA = '"✓,☐,N/A"'    # la ☐ ES la marca y debe seguir imprimiéndose (P3a)
DV_P4 = '"✓,✗,—"'           # la que YA traen los ficheros P4: no se toca
VERDE_OK = 'C8E6C9'         # relleno de la fila completada
ETIQ_CONTADOR = 'Tareas completadas:'

LINEA_MARCA = ('Marca con ✓ en la columna «{}» (desplegable): es la que cuenta '
               'el total de tareas completadas.').format(MARCA_COL)

RX_VERSION = re.compile(r'^Versi[óo]n\s')

# --- §2.2 bio anclada -----------------------------------------------------
# Sustituciones 1:1 en la misma celda. `prefijo=True` = sólo se reemplaza el
# comienzo de la frase y se conserva TODO el resto (el párrafo de proveedores
# de coctelería sigue con su explicación intacta).
BIO_SUBS = [
    dict(prefijo=False,
         old='29 años de experiencia en alta hostelería · 15 años de consultoría gastronómica',
         new='Diseñado por John Guerrero — chef y consultor gastronómico desde 2010, '
             'en cocina desde los 17 años · johnguerrero.es'),
    dict(prefijo=False,
         old='▸ Basada en la experiencia de 15 años de consultoría gastronómica',
         new='▸ Basada en la experiencia de John Guerrero: consultor gastronómico desde 2010, '
             'en cocina desde los 17 años'),
    dict(prefijo=False,
         old='29+ anos de experiencia en hosteleria y 15+ anos en consultoria.',
         new='En cocina desde los 17 años · consultor gastronómico desde 2010.'),
    dict(prefijo=True,
         old='Esta plantilla es el resultado de mas de 29 anos de experiencia en hosteleria, filtrada',
         new='Esta plantilla es el resultado de la experiencia de John Guerrero '
             '(en cocina desde los 17 años, consultor gastronómico desde 2010), filtrada'),
]

#: Regex de «bio vieja» del censo. Se usa SÓLO para verificar que no queda
#: ninguna: si casa algo que no está en BIO_SUBS, la verificación falla y se
#: informa `bio_sin_patron` — nunca se inventa una sustitución.
RX_BIO = re.compile(r'29 a[ñn]os|15 a[ñn]os|a[ñn]os de experiencia|15\+ a[ñn]os|29\+ a[ñn]os')

#: Falso positivo REAL y documentado de RX_BIO: no es una bio, es la vida útil
#: de un horno Josper. Ubicación medida el 2026-08-23 barriendo los 14 productos
#: procesados: `kit-tareas-asador/07-semanales-mensuales.xlsx:Instrucciones:B6`
#: (la nota anterior decía `02-horno-josper-brasas.xlsx`, que es falso: ahí no
#: hay ninguna celda que case el patrón). Se excluye de la verificación en vez
#: de reescribir contenido de negocio.
#: ⚠️ Sincronizado a mano con `FALSOS_POSITIVOS_BIO` de `censo-entregables.py`.
FALSOS_POSITIVOS_BIO = frozenset({
    '▸ El mantenimiento del Josper es clave: un horno bien cuidado dura 15+ anos',
})

#: Alfabetos que nunca deben aparecer (inyecciones de modelo). Igual criterio
#: que `guard_idioma()` de bridge.py.
RX_NONLAT = re.compile('[぀-ヿ㐀-䶿一-鿿가-힯'
                       'Ѐ-ӿ؀-ۿ֐-׿฀-๿]')

#: Una NOTA escrita con «= » delante que Excel interpreta como fórmula
#: (`= CF anuales / MC unitario`). Ninguna fórmula generada empieza por «= »
#: con espacio, así que la firma es fiable. Se informa pero NO se corrige:
#: arreglarla es cambiar texto de negocio, y eso es Fase B (§3).
RX_PSEUDO_FORMULA = re.compile(r'^=\s')

#: Funciones de Excel que pycel NO implementa: su celda se queda sin cache por
#: limitación del evaluador, no porque el fichero esté mal. `PMT` y `COUNTA`
#: vienen documentadas en `inject_cache.py`; `IRR` se comprobó aquí el
#: 2026-08-22 (`kit-plan-financiero/07-informe-viabilidad-bancos.xlsx`,
#: `Proyecciones:B21` → `UnknownFunction`). Se informan aparte y NO tumban la
#: verificación: la fórmula es válida y Excel la calcula al abrir.
FUNCIONES_SIN_PYCEL = ('IRR', 'XIRR', 'PMT', 'COUNTA')
RX_SIN_PYCEL = re.compile(r'\b(' + '|'.join(FUNCIONES_SIN_PYCEL) + r')\s*\(', re.I)

#: §2.6 — fórmula rota con firma INEQUÍVOCA: un `=` de más pegado al nombre de
#: una función justo después de una división (`=$C$4/=AVERAGE(C7,D7)`, hallazgo
#: INT-01 en `kit-escandallos/10-calculadora-pvp.xlsx:Calculadora PVP:E7:E15`).
#: Excel abre esas celdas con `#¿NOMBRE?`. Es la ÚNICA reparación de fórmula que
#: se autoriza en Fase A y sólo con este patrón: `/=` + nombre de función en
#: mayúsculas + `(`. Cualquier otra anomalía de fórmula se informa, no se toca.
RX_FORMULA_ROTA = re.compile(r'/=(?=[A-Z][A-Z0-9_.]*\()')

#: §2.6 — REPARACIONES EXACTAS celda a celda (decisión del orquestador, 2026-08-22,
#: tras el dry-run `all`: 6 productos no pasaban la verificación por REFERENCIAS
#: CIRCULARES preexistentes que Excel abre con el aviso «referencia circular» y
#: pycel no puede cachear). La avería es del generador (filas desplazadas) y la
#: fórmula correcta se deduce sin ambigüedad de las ETIQUETAS de la propia hoja:
#:   · `Margen bruto` = suma de las cuatro filas de arriba (facturación y los
#:     tres costes en negativo), no de sí misma;
#:   · `EBITDA mensual` = margen bruto + costes fijos (en negativo);
#:   · `Margen EBITDA (%)` = EBITDA / facturación bruta;
#:   · `Food cost (€)` se aplica sobre la FACTURACIÓN, no sobre las comisiones
#:     (con la fórmula vieja el food cost salía POSITIVO y sumaba margen);
#:   · `Margen contribución mensual` = facturación × (1 − food cost), no sobre
#:     sí misma.
#: Se aplican SOLO si la celda contiene exactamente la fórmula rota (idempotente
#: y a prueba de ediciones del cliente). Clave: (fichero, hoja, celda).
REPARACIONES_EXACTAS = {
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'B24'): ('=-B23*B7', '=-B22*B7'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'C24'): ('=-C23*B7', '=-C22*B7'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'D24'): ('=-D23*B7', '=-D22*B7'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'B26'): ('=SUM(B26:B29)', '=SUM(B22:B25)'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'C26'): ('=SUM(C26:C29)', '=SUM(C22:C25)'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'D26'): ('=SUM(D26:D29)', '=SUM(D22:D25)'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'B28'): ('=B32+B33', '=B26+B27'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'C28'): ('=C32+C33', '=C26+C27'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'D28'): ('=D32+D33', '=D26+D27'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'B29'): ('=IF(B29=0,0,B35/B29)', '=IF(B22=0,0,B28/B22)'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'C29'): ('=IF(C29=0,0,C35/C29)', '=IF(C22=0,0,C28/C22)'),
    ('calculadora-viabilidad-dark-kitchen.xlsx', 'Punto de Equilibrio', 'D29'): ('=IF(D29=0,0,D35/D29)', '=IF(D22=0,0,D28/D22)'),
    # mismo generador en las 5 guías casual / japonés / mexicano / nikkei / peruano
    ('cash-flow-break-even.xlsx', 'Break-Even', 'B12'): ('=B12*(1-B8)', '=B11*(1-B8)'),
}

HOJAS_TEXTO = ('Instrucciones', 'Índice', 'Indice')

#: §2.1 INT-08 — cadenas que NO son el título del entregable aunque sean la
#: primera celda de texto: rótulos de sección, membretes y marcas. Se comparan
#: sin tildes y en mayúsculas contra el PRINCIPIO de la celda.
DESCARTES_TITULO = ('INSTRUCCIONES', 'INDICE', 'COMO USAR',
                    'AI CHEF PRO —', 'CHEFBUSINESS', 'AICHEF.PRO')

#: Pasos numerados («1. Revisa la hoja…») y viñetas de las instrucciones: tampoco
#: titulan nada (3 planes financieros acababan con su paso 1 como título).
RX_PASO_O_VINETA = re.compile(r'^(\d+[.)]\s|[-–•▸]\s)')

#: Membrete delante de un título bueno («ChefBusiness — Plan Financiero
#: Bar-Restaurante»): se quita el membrete y se conserva el título.
RX_MEMBRETE = re.compile(r'^(ChefBusiness(\.co)?|AI Chef Pro)\s*[—–-]\s*', re.I)

#: Emojis y pictogramas que decoran los títulos (📋, ✅, ▸…): fuera del `title`
#: OOXML, que se lee en un cuadro de diálogo sin soporte de color.
RX_EMOJI = re.compile(
    '[\U0001f000-\U0001faff←-➿⬀-⯿️•▪▸]')


# ==========================================================================
# §2.1 Metadata
# ==========================================================================
def cargar_product_names():
    """`pid` → `productName` leído de `astro-site/src/data/productos/**/<pid>.ts`.

    Es la única fuente del nombre comercial: duplicarlo en el script lo dejaría
    desincronizado en cuanto John renombre un producto.
    """
    nombres = {}
    rx = re.compile(r"productName:\s*\n?\s*(['\"])((?:[^'\"\\]|\\.)*)\1")
    for base, _dirs, ficheros in os.walk(PRODUCTOS_TS):
        for f in ficheros:
            if not f.endswith('.ts') or f == 'types.ts':
                continue
            texto = open(os.path.join(base, f), encoding='utf-8').read()
            m = rx.search(texto)
            if m:
                nombres[f[:-3]] = m.group(2)
    return nombres


def _prefijo_marca(product_name):
    """Parte del `productName` anterior al PRIMER ` — ` (decisión INT-07).

    El corte por `: ` que hacía la versión anterior partía nombres donde los dos
    puntos son parte de la marca (`Kit de Tareas Recurrentes: Tapas Bar /
    Gastrobar`) y dejaba siete productos llamándose todos igual.
    """
    n = (product_name or '').strip()
    return n.split(' — ')[0].strip() if ' — ' in n else n


def calcular_nombres_cortos(nombres):
    """`pid` → nombre corto, resuelto GLOBALMENTE (§2.1, decisiones INT-07/SPC-01).

    El nombre corto es el prefijo de marca, salvo que ese prefijo lo compartan
    dos o más productos (`Kit de Tareas Recurrentes`, `Plan de Negocio`…): en
    ese caso se usa el `productName` COMPLETO, porque un `title` que no
    identifica el producto no sirve de nada. Por eso hace falta ver los 42 a la
    vez y no se puede decidir producto a producto.
    """
    prefijos = {pid: _prefijo_marca(n) for pid, n in nombres.items()}
    repetidos = collections.Counter(p for p in prefijos.values() if p)
    cortos = {}
    for pid, n in nombres.items():
        n = (n or '').strip()
        pref = prefijos[pid]
        if not n:
            cortos[pid] = pid.replace('-', ' ').capitalize()
        elif repetidos[pref] > 1:
            cortos[pid] = n
        else:
            cortos[pid] = pref
    return cortos


def _sin_tildes(s):
    """`s` sin marcas diacríticas, conservando el resto (— incluido)."""
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


def _limpiar_titulo(s):
    """Quita emojis, símbolos iniciales y espacios sobrantes (INT-08)."""
    s = RX_EMOJI.sub('', s)
    s = re.sub(r'^[^0-9A-Za-zÁÉÍÓÚÜÑáéíóúüñ¿¡«"\'(]+', '', s)
    s = RX_MEMBRETE.sub('', s)
    return re.sub(r'\s+', ' ', s).strip()


def _titulo_descartable(s):
    """True si la cadena es un rótulo, un membrete o un paso de las
    instrucciones, no el título (INT-08)."""
    if RX_PASO_O_VINETA.match(s):
        return True
    t = _sin_tildes(s).upper()
    return any(t.startswith(p) for p in DESCARTES_TITULO)


def _primer_titulo_de_hoja(ws):
    """Primera celda de texto de la hoja que sirva como título, o None."""
    for row in ws.iter_rows():
        for c in row:
            if not isinstance(c.value, str):
                continue
            v = _limpiar_titulo(c.value)
            if v and not _titulo_descartable(v):
                return v
    return None


def titulo_legible(wb, fname):
    """Título humano del fichero (§2.1, decisión INT-08).

    Prioridad: hoja de instrucciones/índice → primera hoja de DATOS (su A1, y si
    no vale, su primera celda de texto) → nombre del fichero. En todos los pasos
    se descartan los rótulos (`INSTRUCCIONES`, `ÍNDICE`, `CÓMO USAR`) y los
    membretes (`AI Chef Pro —`, `ChefBusiness…`): no titulan nada, y con la
    versión anterior 30 ficheros acababan llamándose «Instrucciones».
    """
    for ws in wb.worksheets:
        if ws.title in HOJAS_TEXTO:
            t = _primer_titulo_de_hoja(ws)
            if t:
                return t
    for ws in wb.worksheets:
        if ws.title in HOJAS_TEXTO:
            continue
        a1 = ws.cell(row=1, column=1).value
        if isinstance(a1, str):
            v = _limpiar_titulo(a1)
            if v and not _titulo_descartable(v):
                return v
        t = _primer_titulo_de_hoja(ws)
        if t:
            return t
        break   # sólo la PRIMERA hoja de datos, como dice la spec
    base = os.path.splitext(fname)[0]
    base = re.sub(r'^(BONUS-)?\d+[a-z]?-', '', base)
    return base.replace('-', ' ').capitalize()


def set_metadata(wb, pid, fname, corto, cambios):
    """§2.1 — propiedades OOXML de marca. Antes ponía `creator='openpyxl'`,
    que es lo que ve el cliente en Archivo → Información."""
    p = wb.properties
    quiero = dict(
        creator='AI Chef Pro',
        lastModifiedBy='AI Chef Pro',
        title='{} · {}'.format(titulo_legible(wb, fname), corto),
        subject='{} · v{}'.format(corto, VERSION),
        keywords='{}, AI Chef Pro'.format(pid.replace('-', ' ')),
        description='aichef.pro/{}'.format(pid),
        category='AI Chef Pro · Productos digitales',
    )
    for campo, valor in quiero.items():
        if getattr(p, campo) != valor:
            setattr(p, campo, valor)
            cambios.append({'tipo': 'metadata', 'campo': campo, 'valor': valor})


# ==========================================================================
# §2.2 Bio anclada
# ==========================================================================
def aplicar_bio(ws, fname, cambios):
    """§2.2 — sustituye la bio vieja por la anclada, celda a celda.

    Nunca reescribe una celda que no case EXACTAMENTE con uno de los cuatro
    patrones: si aparece una variante nueva, la verificación la denunciará como
    `bio_sin_patron` en vez de que el script improvise.
    """
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            for sub in BIO_SUBS:
                if sub['prefijo']:
                    if v.startswith(sub['old']):
                        c.value = sub['new'] + v[len(sub['old']):]
                        cambios.append({'tipo': 'bio', 'hoja': ws.title,
                                        'celda': c.coordinate, 'modo': 'prefijo'})
                        break
                elif v == sub['old']:
                    c.value = sub['new']
                    cambios.append({'tipo': 'bio', 'hoja': ws.title,
                                    'celda': c.coordinate, 'modo': 'exacto'})
                    break


# ==========================================================================
# §2.6 Reparación de fórmulas inválidas con patrón inequívoco
# ==========================================================================
def reparar_formulas(ws, fname, cambios):
    """§2.6 — repara fórmulas inválidas con firma inequívoca.

    Excepción DOCUMENTADA de la Fase A: son las únicas celdas cuyo texto se toca
    aparte de la bio. Dos vías, ambas idempotentes:
      1. Patrón `/=FUNC(` (hallazgo INT-01): se quita el `=` de más. Sin
         arreglarla la celda no es un defecto estético: Excel la abre con
         `#¿NOMBRE?`.
      2. `REPARACIONES_EXACTAS`: referencias circulares del generador, sólo si
         la celda contiene EXACTAMENTE la fórmula rota conocida.
    No se generaliza a otros operadores ni a otras anomalías: lo demás se
    informa (`pseudo_formulas`) y espera a la Fase B.
    """
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and v.startswith('=') and c.data_type == 'f'):
                continue
            if RX_PSEUDO_FORMULA.match(v):
                # Etiqueta escrita como «= Margen bruto (€)» que el generador guardó
                # como FÓRMULA: Excel la abre con `#¿NOMBRE?`. Mismo texto visible,
                # tipo texto (6 celdas en /dl el 2026-08-22). Idempotente: tras la
                # conversión data_type es 's' y no vuelve a entrar aquí.
                c.data_type = 's'
                cambios.append({'tipo': 'pseudo_formula_a_texto', 'hoja': ws.title,
                                'celda': c.coordinate, 'texto': v})
                continue
            nuevo = v
            exacta = REPARACIONES_EXACTAS.get((fname, ws.title, c.coordinate))
            if exacta and v == exacta[0]:
                nuevo = exacta[1]
            elif '/=' in v:
                nuevo = RX_FORMULA_ROTA.sub('/', v)
            if nuevo != v:
                c.value = nuevo
                cambios.append({'tipo': 'formula_reparada', 'hoja': ws.title,
                                'celda': c.coordinate, 'antes': v, 'despues': nuevo})


# ==========================================================================
# §2.3 Impresión A4
# ==========================================================================
def print_setup(ws, header_row, landscape, cambios):
    """§2.3 — A4 ajustado al ancho, márgenes estrechos y pie de marca.

    SÓLO se llama cuando `paperSize is None`: si la hoja ya venía configurada
    (los 08/09 de los kits, p. ej.) se respeta tal cual, porque alguien la
    ajustó a mano y no hay forma de saber por qué.

    La ORIENTACIÓN sigue el mismo criterio (decisión INT-02): la heurística de
    «≥ 6 columnas → apaisado» sólo decide cuando la hoja no traía orientación.
    Si venía fijada se conserva y se anota — tres hojas estrechas de pack-appcc
    y del BONUS-02 de kit-tareas están en apaisado a propósito.
    """
    orientacion_previa = ws.page_setup.orientation
    ws.page_setup.paperSize = 9  # A4
    if not orientacion_previa:
        ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.59, right=0.59, top=0.59, bottom=0.59,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = 'AI Chef Pro · aichef.pro · Página &P de &N'
    ws.oddFooter.center.size = 8
    detalle = {'tipo': 'impresion', 'hoja': ws.title,
               'orientacion': ws.page_setup.orientation}
    if orientacion_previa:
        detalle['orientacion_conservada'] = orientacion_previa
    if header_row:
        ws.print_title_rows = '{0}:{0}'.format(header_row)
        detalle['print_title_rows'] = header_row
        # El freeze existente manda: si la hoja ya congelaba algo, quien la hizo
        # sabía dónde quería el corte.
        if ws.freeze_panes is None:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
            detalle['freeze_panes'] = ws.freeze_panes
    cambios.append(detalle)


def configurar_impresion(ws, hr, cambios):
    """Decide orientación y títulos según el tipo de hoja (§2.3)."""
    if ws.page_setup.paperSize is not None:
        return
    es_texto = ws.title in HOJAS_TEXTO or (ws.max_column <= 3 and hr is None)
    if es_texto:
        print_setup(ws, None, landscape=False, cambios=cambios)
    else:
        print_setup(ws, hr, landscape=ws.max_column >= 6, cambios=cambios)


# ==========================================================================
# §2.4 Casilla unificada — geometría común
# ==========================================================================
def fila_cabecera(ws):
    """Primera fila 1..6 con ≥ 3 celdas de texto, NINGUNA numérica y fila
    siguiente con contenido (§2.4, endurecido por la decisión INT-05).

    Es el criterio de §2.4. Deliberadamente estructural (y no «la fila que dice
    Tarea»): los cinco patrones usan encabezados distintos y en tres de ellos la
    palabra «Tarea» ni aparece.

    Las dos restricciones nuevas —tope en la fila 6 y cero celdas numéricas—
    salen de un falso positivo medido: en
    `plan-negocio-cocteleria-eventos/calculadora-pricing-eventos.xlsx:Calculadora`
    la fila 7 es `Tipo de evento | Boda | … | 6,5`, un renglón de DATOS, y se
    estaba congelando y repitiendo en cada página como si fuera la cabecera.
    Una cabecera de verdad no tiene números.
    """
    ncol = min(ws.max_column, 40)
    for r in range(1, min(7, ws.max_row + 1)):
        textos = 0
        numerica = False
        for c in range(1, ncol + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                textos += 1
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                numerica = True
        if numerica or textos < 3:
            continue
        siguiente = any(ws.cell(row=r + 1, column=c).value is not None
                        for c in range(1, ncol + 1))
        if siguiente:
            return r
    return None


def _es_fila_seccion(ws, fila):
    """True si la fila es un separador de sección (celda combinada desde A)."""
    for m in ws.merged_cells.ranges:
        if m.min_row == fila and m.max_row == fila and m.min_col == 1 and m.max_col >= 3:
            return True
    return False


def _fila_libre(ws, fila):
    """True si la fila está vacía y sin combinaciones: se puede escribir en ella
    sin desplazar nada (§2.4 prohíbe insertar filas)."""
    if any(m.min_row <= fila <= m.max_row for m in ws.merged_cells.ranges):
        return False
    return all(ws.cell(row=fila, column=c).value is None
               for c in range(1, ws.max_column + 1))


def _fila_para_contador(ws, fin):
    """Fila donde cabe un contador NUEVO (decisión INT-09).

    Primero `última + 2` (deja una fila de aire, que es como se ven los kits ya
    buenos) y, si no está libre o está combinada, `última + 1`. Si tampoco,
    None: §3 prohíbe insertar filas, así que la hoja se queda sin contador y
    queda anotada en el informe.
    """
    for candidata in (fin + 2, fin + 1):
        if _fila_libre(ws, candidata):
            return candidata
    return None


def _contar_texto_col(ws, col, desde, hasta):
    """Equivalente a `COUNTIF(<col>{desde}:<col>{hasta},"?*")`.

    `"?*"` en Excel cuenta celdas de TEXTO de al menos un carácter: los números
    no entran. Se evalúa aquí para poder decidir sin abrir pycel.
    """
    n = 0
    for r in range(desde, hasta + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            n += 1
    return n


def _contar_numeradas(ws, desde, hasta):
    """Filas con un entero en la columna A: las tareas numeradas de P1/P2."""
    return sum(1 for r in range(desde, hasta + 1)
               if isinstance(ws.cell(row=r, column=1).value, int)
               and not isinstance(ws.cell(row=r, column=1).value, bool))


def _buscar_fila(ws, texto):
    """Fila que contiene exactamente `texto` en cualquier columna (o None)."""
    for row in ws.iter_rows():
        for c in row:
            if c.value == texto:
                return c.row
    return None


def _col_por_cabecera(ws, hr, nombres):
    """Columna cuya cabecera coincide con alguno de `nombres` (la última gana,
    igual que en el post-proceso de pastelería)."""
    encontrada = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hr, column=c).value
        if isinstance(v, str) and v.strip() in nombres:
            encontrada = c
    return encontrada


def _dv_lista_de_col(ws, col):
    """Validaciones de lista que tocan la columna `col`."""
    out = []
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        if any(r.min_col <= col <= r.max_col for r in dv.sqref.ranges):
            out.append(dv)
    return out


def _coords(dv):
    """Conjunto de coordenadas (fila, col) cubiertas por una validación."""
    s = set()
    for r in dv.sqref.ranges:
        s |= set(r.cells)
    return s


def poner_dv(ws, col, filas, formula, cambios, motivo):
    """Deja la columna `col` con UNA validación de lista `formula` sobre `filas`.

    Idempotente: si ya existe exactamente esa validación sobre exactamente esas
    celdas, no toca nada y no cuenta cambio.
    """
    if not filas:
        return
    objetivo = {(r, col) for r in filas}
    existentes = _dv_lista_de_col(ws, col)
    if len(existentes) == 1 and existentes[0].formula1 == formula \
            and _coords(existentes[0]) == objetivo:
        return
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if dv not in existentes]
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    letra = get_column_letter(col)
    for r in sorted(filas):
        dv.add(ws.cell(row=r, column=col))
    cambios.append({'tipo': 'dv', 'hoja': ws.title, 'columna': letra,
                    'formula': formula, 'celdas': len(filas), 'motivo': motivo})


def poner_cf_fila_verde(ws, col_marca, hr, fin, ncols, cambios):
    """Formato condicional: la fila entera se pone verde cuando la marca es ✓.

    Se borran antes las reglas cuya fórmula mencione `"✓"` para no acumular
    duplicados en cada pasada; si la regla que queda ya es la deseada, se sale
    sin tocar nada (idempotencia).
    """
    if fin <= hr:
        return
    letra = get_column_letter(col_marca)
    rango = 'A{}:{}{}'.format(hr + 1, get_column_letter(ncols), fin)
    formula = '${}{}="✓"'.format(letra, hr + 1)

    reglas = ws.conditional_formatting._cf_rules
    ya = False
    sobra = False
    for k, rs in list(reglas.items()):
        for r in rs:
            if any('"✓"' in str(f) for f in (getattr(r, 'formula', None) or [])):
                if str(k.sqref) == rango and r.formula and r.formula[0] == formula:
                    ya = True
                else:
                    sobra = True
    if ya and not sobra:
        return

    nuevas = type(reglas)()
    for k, rs in reglas.items():
        quedan = [r for r in rs
                  if not any('"✓"' in str(f) for f in (getattr(r, 'formula', None) or []))]
        if quedan:
            nuevas[k] = quedan
    ws.conditional_formatting._cf_rules = nuevas
    ws.conditional_formatting.add(rango, FormulaRule(
        formula=[formula],
        fill=PatternFill('solid', start_color=VERDE_OK, end_color=VERDE_OK)))
    cambios.append({'tipo': 'cf', 'hoja': ws.title, 'rango': rango, 'formula': formula})


def escribir_contador(ws, fila, col_label, col_num, col_marca, hr, fin, cambios,
                      nuevo, informe_hoja=None):
    """Contador «Tareas completadas: N de M» con AMBOS lados calculados.

    El denominador se calcula sobre la columna B (`COUNTIF(B…,"?*")` cuenta
    celdas con texto): si el usuario borra o añade tareas, el total deja de
    mentir. Las filas de sección van con la columna B vacía, así que no cuentan.

    EXCEPCIÓN (decisión INT-03) — las plantillas EN BLANCO. En
    `kit-tareas/07-plantilla-personalizable.xlsx` hay 15 filas numeradas y sólo
    3 con texto en B («(Escribe aquí tu tarea)»); el denominador calculado diría
    «0 de 3» sobre una plantilla pensada para 15 tareas. Si el denominador
    original es un número ESTÁTICO y el COUNTIF evaluado es menor que las filas
    numeradas, se conserva el número original y se anota.
    """
    letra = get_column_letter(col_marca)
    den_formula = '=COUNTIF(B{}:B{},"?*")'.format(hr + 1, fin)
    den_actual = ws.cell(row=fila, column=col_num + 2).value
    if (not nuevo and isinstance(den_actual, (int, float))
            and not isinstance(den_actual, bool)
            and _contar_texto_col(ws, 2, hr + 1, fin) < _contar_numeradas(ws, hr + 1, fin)):
        den_formula = den_actual
        if informe_hoja is not None:
            informe_hoja['nota'] = ('denominador estático conservado (plantilla): '
                                    '{} = {!r}'.format(
                                        ws.cell(row=fila, column=col_num + 2).coordinate,
                                        den_actual))
    valores = [
        (col_label, ETIQ_CONTADOR),
        (col_num, '=COUNTIF({0}{1}:{0}{2},"✓")'.format(letra, hr + 1, fin)),
        (col_num + 1, 'de'),
        (col_num + 2, den_formula),
    ]
    tocado = False
    for col, valor in valores:
        cel = ws.cell(row=fila, column=col)
        if cel.value != valor:
            cel.value = valor
            tocado = True
        if col == col_label and (cel.font is None or not cel.font.bold):
            cel.font = Font(name='Calibri', size=10, bold=True)
    if tocado:
        cambios.append({'tipo': 'contador', 'hoja': ws.title,
                        'celda': ws.cell(row=fila, column=col_num).coordinate,
                        'fila': fila, 'nuevo': nuevo,
                        'rango': '{0}{1}:{0}{2}'.format(letra, hr + 1, fin)})


def _col_num_del_contador(ws, fila, col_defecto):
    """Columna del numerador en un contador ya existente: la anterior al «de»."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=fila, column=c).value == 'de':
            return c - 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila, column=c).value
        if isinstance(v, str) and v.startswith('=COUNTIF'):
            return c
    return col_defecto


# ==========================================================================
# §2.4 Detección de patrón
# ==========================================================================
def detectar_patron(ws):
    """Clasifica la hoja en P1..P5 (§2.4) y devuelve su geometría.

    Se decide por la CABECERA, no por el producto: el mismo kit mezcla hojas de
    tres patrones distintos (los 08/09 son P4 aunque el 01 sea P1).
    """
    hr = fila_cabecera(ws)
    if hr is None:
        return 'P5', {'motivo': 'sin fila de cabecera'}
    a = ws.cell(row=hr, column=1).value
    b = ws.cell(row=hr, column=2).value
    a = a.strip() if isinstance(a, str) else a
    b = b.strip() if isinstance(b, str) else b

    filas_casilla = [r for r in range(hr + 1, ws.max_row + 1)
                     if ws.cell(row=r, column=1).value == '☐']
    contador = _buscar_fila(ws, ETIQ_CONTADOR)

    # --- P1 / P2: familia «▸», ☐ decorativa en A y columna de marca aparte ---
    if a in ('☐', 'Nº') and b == 'Tarea':
        col_marca = _col_por_cabecera(ws, hr, ('Hecha', MARCA_COL))
        if col_marca:
            geom = {'hr': hr, 'col_marca': col_marca, 'contador': contador,
                    'filas_casilla': filas_casilla}
            return ('P1' if contador else 'P2'), geom

    # --- P3a: la ☐ de la columna A ES la marca ------------------------------
    # No se filtra por el nombre de la columna B: `checklist-contratacion` de
    # la guía de panadería la llama «Tarea» igual que los P1, y sin embargo no
    # tiene columna de marca separada — la ☐ de A es lo único marcable.
    if (a in ('☐', 'OK', '✓') or a is None) and filas_casilla:
        return 'P3a', {'hr': hr, 'col_marca': 1, 'contador': contador,
                       'filas_casilla': filas_casilla}

    # --- P4: numeración + desplegable "✓,✗,—" ya montado --------------------
    col_marca = _col_por_cabecera(ws, hr, ('✓', MARCA_COL, 'Hecha'))
    if col_marca and a in ('#', 'Nº'):
        dvs = _dv_lista_de_col(ws, col_marca)
        if any(dv.formula1 == DV_P4 for dv in dvs):
            return 'P4', {'hr': hr, 'col_marca': col_marca, 'contador': contador}

    # --- P3b: columna «OK» fuera de A, estilo ChefBusiness ------------------
    # La condición «sin DV» valía para la PRIMERA pasada y rompía la
    # idempotencia en la segunda: en cuanto este script le pone su propio
    # desplegable, la hoja dejaba de reconocerse como P3b y caía a P5 (decisión
    # INT-06). Se acepta también la DV que pone este script — y sólo esa.
    col_ok = _col_por_cabecera(ws, hr, ('OK',))
    if col_ok and col_ok > 1:
        dvs = _dv_lista_de_col(ws, col_ok)
        if not dvs or all(dv.formula1 == DV_TAREA for dv in dvs):
            return 'P3b', {'hr': hr, 'col_marca': col_ok, 'contador': contador}

    return 'P5', {'motivo': 'cabecera sin columna de marca reconocible', 'hr': hr}


# ==========================================================================
# §2.4 Normalizadores por patrón
# ==========================================================================
def _contador_nuevo(ws, hr, fin, col_label, col_num, col_marca, cambios, informe_hoja):
    """Coloca un contador NUEVO si procede (decisiones INT-04 e INT-09).

    Dos motivos para no ponerlo, los dos anotados en el informe: que la hoja sea
    una plantilla en blanco (denominador 0 — un «0 de 0» no informa de nada y
    encima parece un error) o que no haya una fila libre donde escribirlo.
    """
    if _contar_texto_col(ws, 2, hr + 1, fin) == 0:
        informe_hoja['nota'] = 'sin contador: plantilla sin tareas'
        return
    if col_num + 2 > ws.max_column:
        informe_hoja['nota'] = (
            'sin contador: no caben numerador/de/denominador a partir de la '
            'columna {}'.format(get_column_letter(col_num)))
        return
    fila = _fila_para_contador(ws, fin)
    if fila is None:
        informe_hoja['nota'] = ('sin contador: ni la fila {} ni la {} están '
                                'libres'.format(fin + 2, fin + 1))
        return
    escribir_contador(ws, fila, col_label, col_num, col_marca, hr, fin, cambios,
                      nuevo=True, informe_hoja=informe_hoja)


def normalizar_P1(ws, geom, cambios, informe_hoja):
    """P1 — «▸» con contador (kit-tareas base y sus 6 clones, 01-07 + BONUS-01).

    La ☐ de la columna A no era marcable (no tenía desplegable): era decorativa
    y encima duplicaba la columna «Hecha», que es la que cuenta el contador.
    Pasa a ser el número de tarea, y la marca real se llama y se comporta igual
    que en pastelería v2.0.
    """
    hr, cm, contador = geom['hr'], geom['col_marca'], geom['contador']
    ncols = ws.max_column
    fin = contador - 2  # las filas de holgura ya existen; no se insertan

    _renumerar_col_a(ws, hr, cambios)
    _renombrar_cabecera_marca(ws, hr, cm, cambios)

    ultima = _ultima_tarea(ws, hr, fin)
    filas_dv = [r for r in range(hr + 1, fin + 1)
                if isinstance(ws.cell(row=r, column=1).value, int)
                or (r > ultima and _fila_libre(ws, r))]
    poner_dv(ws, cm, filas_dv, DV_TAREA, cambios, 'marca de tarea')
    poner_cf_fila_verde(ws, cm, hr, fin, ncols, cambios)

    col_num = _col_num_del_contador(ws, contador, 4)
    escribir_contador(ws, contador, 1, col_num, cm, hr, fin, cambios,
                      nuevo=False, informe_hoja=informe_hoja)
    return fin


def normalizar_P2(ws, geom, cambios, informe_hoja):
    """P2 — «▸» sin contador (asador y sushi-bar, 01-09).

    Mismo molde que P1 pero los ficheros salieron sin desplegable ni fórmulas:
    hay que ponerlos, y crear el contador si cabe SIN desplazar nada.
    """
    hr, cm = geom['hr'], geom['col_marca']
    ncols = ws.max_column

    _renumerar_col_a(ws, hr, cambios)
    _renombrar_cabecera_marca(ws, hr, cm, cambios)

    fin = _ultima_tarea(ws, hr, ws.max_row)
    if fin is None:
        informe_hoja['nota'] = 'P2 sin filas de tarea'
        return None
    filas_dv = [r for r in range(hr + 1, fin + 1)
                if isinstance(ws.cell(row=r, column=1).value, int)]
    poner_dv(ws, cm, filas_dv, DV_TAREA, cambios, 'marca de tarea')
    poner_cf_fila_verde(ws, cm, hr, fin, ncols, cambios)

    _contador_nuevo(ws, hr, fin, 1, 4, cm, cambios, informe_hoja)
    return fin


def normalizar_P3a(ws, geom, cambios, informe_hoja, con_contador=True):
    """P3a — la ☐ de la columna A es la marca (checklists de las guías y planes).

    Aquí NO se renumera: si se sustituye la ☐ por un número, el checklist deja
    de poder imprimirse y marcarse a boli, que es como se usa en una apertura.
    La ☐ se queda como valor y recibe un desplegable «✓,☐,N/A» para poder
    marcarla también en pantalla.
    """
    hr, filas = geom['hr'], geom['filas_casilla']
    ncols = ws.max_column
    cab = ws.cell(row=hr, column=1)
    if cab.value == '☐':
        cab.value = '✓'
        cambios.append({'tipo': 'cabecera', 'hoja': ws.title,
                        'celda': cab.coordinate, 'valor': '✓'})
    fin = max(filas)
    poner_dv(ws, 1, filas, DV_CASILLA, cambios, 'casilla marcable')
    poner_cf_fila_verde(ws, 1, hr, fin, ncols, cambios)

    if not con_contador:
        informe_hoja['nota'] = 'sin contador por decisión de producto (no son tareas)'
        return fin
    contador = geom['contador']
    if contador is None:
        _contador_nuevo(ws, hr, fin, 2, 3, 1, cambios, informe_hoja)
    else:
        col_num = _col_num_del_contador(ws, contador, 3)
        escribir_contador(ws, contador, 2, col_num, 1, hr, fin, cambios,
                          nuevo=False, informe_hoja=informe_hoja)
    return fin


def normalizar_P3b(ws, geom, cambios, informe_hoja):
    """P3b — columna «OK» fuera de A, estilo ChefBusiness (kits y planes de origen CB).

    La columna OK venía vacía y sin desplegable: nada indicaba cómo marcar.
    NO se toca el membrete «ChefBusiness Consultoria Gastronomica» de A1
    (decisión pendiente de John, §3 de la spec).
    """
    hr, cm = geom['hr'], geom['col_marca']
    ncols = ws.max_column
    fin = None
    for r in range(hr + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip() and not _es_fila_seccion(ws, r):
            fin = r
    if fin is None:
        informe_hoja['nota'] = 'P3b sin filas de tarea'
        return None
    filas = [r for r in range(hr + 1, fin + 1)
             if isinstance(ws.cell(row=r, column=2).value, str)
             and ws.cell(row=r, column=2).value.strip()
             and not _es_fila_seccion(ws, r)]
    poner_dv(ws, cm, filas, DV_TAREA, cambios, 'marca de tarea')
    poner_cf_fila_verde(ws, cm, hr, fin, ncols, cambios)

    contador = geom['contador']
    if contador is None:
        _contador_nuevo(ws, hr, fin, 1, cm, cm, cambios, informe_hoja)
    else:
        col_num = _col_num_del_contador(ws, contador, cm)
        escribir_contador(ws, contador, 1, col_num, cm, hr, fin, cambios,
                          nuevo=False, informe_hoja=informe_hoja)
    return fin


def normalizar_P4(ws, geom, cambios, informe_hoja):
    """P4 — ya trae Nº y desplegable «✓,✗,—»: sólo le faltaba el feedback visual.

    Su desplegable NO se toca (§2.4): tiene un valor ✗ que significa «no se ha
    podido hacer», distinto de «—», y cambiarlo perdería información.
    """
    hr, cm = geom['hr'], geom['col_marca']
    ncols = ws.max_column
    fin = _ultima_tarea(ws, hr, ws.max_row)
    if fin is None:
        informe_hoja['nota'] = 'P4 sin filas numeradas'
        return None
    poner_cf_fila_verde(ws, cm, hr, fin, ncols, cambios)

    contador = geom['contador']
    if contador is None:
        _contador_nuevo(ws, hr, fin, 2, 3, cm, cambios, informe_hoja)
    else:
        col_num = _col_num_del_contador(ws, contador, 3)
        escribir_contador(ws, contador, 2, col_num, cm, hr, fin, cambios,
                          nuevo=False, informe_hoja=informe_hoja)
    return fin


def _renumerar_col_a(ws, hr, cambios):
    """Cabecera A → «Nº» y cada ☐ decorativa → su número correlativo."""
    cab = ws.cell(row=hr, column=1)
    if cab.value == '☐':
        cab.value = 'Nº'
        cambios.append({'tipo': 'cabecera', 'hoja': ws.title,
                        'celda': cab.coordinate, 'valor': 'Nº'})
    n = 0
    marcadas = []
    for r in range(hr + 1, ws.max_row + 1):
        c = ws.cell(row=r, column=1)
        if c.value == '☐':
            n += 1
            c.value = n
            c.font = Font(name='Calibri', size=10, color='666666')
            c.alignment = Alignment(horizontal='center', vertical='center')
            marcadas.append(c.coordinate)
    if marcadas:
        cambios.append({'tipo': 'numeracion', 'hoja': ws.title,
                        'celdas': len(marcadas), 'desde': marcadas[0],
                        'hasta': marcadas[-1]})


def _renombrar_cabecera_marca(ws, hr, cm, cambios):
    """«Hecha» → «✓ Completada» (vocabulario único de pastelería v2.0)."""
    cel = ws.cell(row=hr, column=cm)
    if cel.value != MARCA_COL:
        cel.value = MARCA_COL
        cambios.append({'tipo': 'cabecera', 'hoja': ws.title,
                        'celda': cel.coordinate, 'valor': MARCA_COL})


def _ultima_tarea(ws, hr, tope):
    """Última fila (≤ tope) con un entero en la columna A: la última tarea."""
    ultima = None
    for r in range(hr + 1, min(tope, ws.max_row) + 1):
        if isinstance(ws.cell(row=r, column=1).value, int):
            ultima = r
    return ultima


# ==========================================================================
# §2.5 Línea de versión y de marca
# ==========================================================================
def linea_version(ws, texto, cambios, rx=None):
    """Escribe `texto` en la hoja de instrucciones: sustituye la línea que case
    con `rx` o la añade 2 filas por debajo de la última de texto, copiando su
    estilo. Nunca duplica (idempotente)."""
    col = 2 if ws.cell(row=2, column=2).value else 1
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str):
            if v == texto:
                return
            if rx and rx.match(v):
                ws.cell(row=r, column=col).value = texto
                cambios.append({'tipo': 'version', 'hoja': ws.title,
                                'celda': ws.cell(row=r, column=col).coordinate,
                                'modo': 'sustituida'})
                return
    origen = None
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(row=r, column=col).value, str):
            origen = r
            break
    destino = ws.max_row + 2
    cel = ws.cell(row=destino, column=col, value=texto)
    if origen:
        cel._style = copy.copy(ws.cell(row=origen, column=col)._style)
    cambios.append({'tipo': 'version', 'hoja': ws.title,
                    'celda': cel.coordinate, 'modo': 'añadida'})


# ==========================================================================
# §2.6 Limpieza, cierre y verificación
# ==========================================================================
def limpiar_vacios(wb, cambios):
    """Celdas con cadena vacía `''` → None.

    Una celda con `''` produce XML no conforme (`<c t="inlineStr"/>` sin valor)
    y algunos visores la pintan como error. Auditoría L1-14 de pastelería.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value == '' and not isinstance(c.value, (int, float)):
                    c.value = None
                    cambios.append({'tipo': 'vacio', 'hoja': ws.title,
                                    'celda': c.coordinate})


def verificar(path, skip_cache=False):
    """§2.6 — vuelve a abrir el fichero y comprueba las 8 condiciones de cierre.

    Se abre DOS veces: con fórmulas (para inspeccionar el texto) y con
    `data_only` (para leer el valor cacheado que dejó `inject_cache.py`). Si el
    fichero no vuelve a abrir, el XML se rompió y se informa como `abre=False`.
    """
    v = {'abre': True, 'creator': None, 'formulas': 0, 'nocache_total': 0,
         'nocache_vacio': 0, 'nocache_real': 0, 'nonlat': 0, 'bio_vieja': [],
         'bio_falso_positivo': [], 'empty_str': 0, 'noprint': 0,
         'contadores': 0, 'contadores_ko': [], 'pseudo_formulas': [],
         'nocache_pycel': [], 'ok': False}
    try:
        wbf = openpyxl.load_workbook(path)
        wbv = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:                                   # XML roto
        v['abre'] = False
        v['error'] = '{}: {}'.format(type(e).__name__, e)
        return v

    v['creator'] = wbf.properties.creator
    filas_contador = {}
    for ws, wsv in zip(wbf.worksheets, wbv.worksheets):
        if ws.page_setup.paperSize != 9:
            v['noprint'] += 1
        for row in ws.iter_rows():
            for c in row:
                val = c.value
                if val == '' and not isinstance(val, (int, float)):
                    v['empty_str'] += 1
                if not isinstance(val, str):
                    continue
                if RX_NONLAT.search(val):
                    v['nonlat'] += 1
                if RX_BIO.search(val):
                    destino = 'bio_falso_positivo' if val in FALSOS_POSITIVOS_BIO else 'bio_vieja'
                    v[destino].append('{}:{}'.format(ws.title, c.coordinate))
                if val == ETIQ_CONTADOR:
                    filas_contador.setdefault(ws.title, set()).add(c.row)
                if c.data_type != 'f':
                    continue   # texto (incluidas las etiquetas «= …» ya convertidas a texto)
                if RX_PSEUDO_FORMULA.match(val):
                    # nota disfrazada de fórmula que reparar_formulas() no convirtió:
                    # se denuncia con fichero:hoja:celda (no debería quedar ninguna)
                    v['pseudo_formulas'].append('{}:{} = {!r}'.format(
                        ws.title, c.coordinate, val))
                else:
                    v['formulas'] += 1
                    if wsv[c.coordinate].value is None:
                        v['nocache_total'] += 1
                        if '""' in val:
                            v['nocache_vacio'] += 1   # IF(x="","",…): vacía por diseño
                        elif RX_SIN_PYCEL.search(val):
                            v['nocache_pycel'].append('{}:{} = {!r}'.format(
                                ws.title, c.coordinate, val))
                        else:
                            v['nocache_real'] += 1

    # las fórmulas del contador tienen que EVALUAR (§2.6)
    for ws, wsv in zip(wbf.worksheets, wbv.worksheets):
        for fila in sorted(filas_contador.get(ws.title, ())):
            for c in range(1, ws.max_column + 1):
                f = ws.cell(row=fila, column=c).value
                if not (isinstance(f, str) and f.startswith('=COUNTIF')):
                    continue
                v['contadores'] += 1
                got = wsv.cell(row=fila, column=c).value
                if not isinstance(got, (int, float)) or isinstance(got, bool) or got < 0:
                    v['contadores_ko'].append('{}:{}={!r}'.format(
                        ws.title, ws.cell(row=fila, column=c).coordinate, got))

    # `contadores_ko` se exime con --skip-cache por la misma razón que
    # `nocache_real` (decisión SPC-04): sin `inject_cache.py` NINGUNA fórmula
    # tiene valor cacheado, así que las del contador salen todas en rojo y el
    # fallo sería un artefacto del modo, no del fichero.
    v['ok'] = (v['abre'] and v['creator'] == 'AI Chef Pro' and v['nonlat'] == 0
               and not v['bio_vieja'] and v['empty_str'] == 0 and v['noprint'] == 0
               and (skip_cache or (v['nocache_real'] == 0 and not v['contadores_ko'])))
    return v


# ==========================================================================
# Proceso por fichero y por producto
# ==========================================================================
#: Hojas que NO llevan contador aunque encajen en P3a: no son tareas.
SIN_CONTADOR = {('plan-negocio-cocteleria-eventos',
                 'plantilla-proveedores-cocteleria-eventos.xlsx')}


def procesar_fichero(path, pid, corto, informe):
    """Aplica §2.1-§2.6 (menos el cache) a un .xlsx y devuelve sus cambios."""
    fname = os.path.basename(path)
    cambios = []
    hojas = {}
    wb = openpyxl.load_workbook(path)

    set_metadata(wb, pid, fname, corto, cambios)
    for ws in wb.worksheets:
        aplicar_bio(ws, fname, cambios)
        reparar_formulas(ws, fname, cambios)

    hay_p1_p2 = False
    for ws in wb.worksheets:
        patron, geom = detectar_patron(ws)
        info = {'patron': patron}
        if patron == 'P1':
            normalizar_P1(ws, geom, cambios, info)
            hay_p1_p2 = True
        elif patron == 'P2':
            normalizar_P2(ws, geom, cambios, info)
            hay_p1_p2 = True
        elif patron == 'P3a':
            normalizar_P3a(ws, geom, cambios, info,
                           con_contador=(pid, fname) not in SIN_CONTADOR)
        elif patron == 'P3b':
            normalizar_P3b(ws, geom, cambios, info)
        elif patron == 'P4':
            normalizar_P4(ws, geom, cambios, info)
        else:
            info['motivo'] = geom.get('motivo')
        hojas[ws.title] = info

    # §2.5 — versión y, en los P1/P2, cómo se marca la casilla
    texto_version = 'Versión {} · {} · aichef.pro/{} · info@aichef.pro'.format(
        VERSION, MES, pid)
    for ws in wb.worksheets:
        if ws.title in HOJAS_TEXTO:
            if hay_p1_p2:
                linea_version(ws, LINEA_MARCA, cambios)
            linea_version(ws, texto_version, cambios, RX_VERSION)

    limpiar_vacios(wb, cambios)

    # §2.3 — impresión al final, cuando ya se conocen filas y columnas finales
    for ws in wb.worksheets:
        configurar_impresion(ws, fila_cabecera(ws), cambios)

    # Sin cambios NO se guarda: un `save()` de openpyxl borraría el cache de
    # valores que dejó inject_cache.py y obligaría a reinyectarlo. Así la
    # segunda pasada no toca el fichero ni por dentro ni en el mtime.
    if cambios:
        wb.save(path)
    informe['hojas'] = hojas
    informe['cambios'] = cambios
    return cambios


def resumen_cambios(cambios):
    """Recuento por tipo, que es lo que se mira para juzgar una pasada."""
    out = {}
    for c in cambios:
        out[c['tipo']] = out.get(c['tipo'], 0) + 1
    return out


def procesar_producto(carpeta, pid, corto, skip_cache):
    """Procesa todos los .xlsx de una carpeta, inyecta el cache y verifica."""
    ficheros = sorted(f for f in os.listdir(carpeta)
                      if f.endswith('.xlsx') and not f.startswith('~$'))
    informe = {'producto': pid, 'carpeta': carpeta, 'nombre_corto': corto,
               'ficheros': {}}
    reescritos = []
    for f in ficheros:
        informe['ficheros'][f] = {}
        if procesar_fichero(os.path.join(carpeta, f), pid, corto, informe['ficheros'][f]):
            reescritos.append(f)

    # El cache va SIEMPRE al final: cualquier wb.save() posterior lo borra.
    # Sólo sobre los ficheros reescritos — los intactos ya lo conservan.
    if not skip_cache and reescritos:
        subprocess.run([sys.executable, INJECT_CACHE]
                       + [os.path.join(carpeta, f) for f in reescritos],
                       check=True, stdout=subprocess.DEVNULL)

    ok = True
    for f in ficheros:
        v = verificar(os.path.join(carpeta, f), skip_cache=skip_cache)
        informe['ficheros'][f]['verificacion'] = v
        informe['ficheros'][f]['resumen'] = resumen_cambios(
            informe['ficheros'][f]['cambios'])
        ok &= v['ok']
    informe['total_cambios'] = resumen_cambios(
        [c for f in ficheros for c in informe['ficheros'][f]['cambios']])
    informe['patrones'] = {}
    # Decisión INT-09: hay que poder ver de un vistazo cuántas hojas con tareas
    # se quedaron sin contador y por qué, sin bucear en el JSON hoja a hoja.
    informe['sin_contador'] = []
    for f in ficheros:
        for hoja, info in informe['ficheros'][f]['hojas'].items():
            informe['patrones'][info['patron']] = \
                informe['patrones'].get(info['patron'], 0) + 1
            nota = info.get('nota', '')
            if nota.startswith('sin contador') and info['patron'] in (
                    'P2', 'P3a', 'P3b', 'P4'):
                informe['sin_contador'].append('{}:{} — {}'.format(f, hoja, nota))
    informe['ok'] = ok
    return informe


# ==========================================================================
# CLI
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('objetivo', nargs='?',
                    help='productId, "all" o una ruta de carpeta')
    ap.add_argument('--dry-run', action='store_true',
                    help='copia a $SCRATCH/dryrun/<pid>/ y procesa allí')
    ap.add_argument('--json', dest='json_out', help='ruta del informe JSON')
    ap.add_argument('--skip-cache', action='store_true',
                    help='no ejecutar inject_cache.py')
    ap.add_argument('--nombres-cortos', action='store_true',
                    help='imprime el nombre corto de cada producto y termina')
    args = ap.parse_args()

    nombres = cargar_product_names()
    cortos = calcular_nombres_cortos(nombres)

    if args.nombres_cortos:
        for pid in sorted(cortos):
            print('{:42s} {}'.format(pid, cortos[pid]))
        distintos = len(set(cortos.values()))
        print('{} productos · {} nombres cortos distintos · {}'.format(
            len(cortos), distintos,
            'TODOS ÚNICOS' if distintos == len(cortos) else 'HAY COLISIONES'))
        return 0 if distintos == len(cortos) else 1

    if not args.objetivo:
        ap.error('falta el objetivo (productId, "all" o una ruta de carpeta)')

    objetivos = []   # (pid, carpeta_origen)
    if args.objetivo == 'all':
        for d in sorted(os.listdir(DL)):
            ruta = os.path.join(DL, d)
            if os.path.isdir(ruta) and d not in EXCLUIDOS:
                objetivos.append((d, ruta))
    elif os.path.isdir(args.objetivo):
        ruta = os.path.abspath(args.objetivo).rstrip(os.sep)
        pid = os.path.basename(ruta)
        # Decisión SPC-03: el guard de EXCLUIDOS también vale por ruta cuando la
        # carpeta está DENTRO de /dl. Antes, `… /dl/kit-tareas-pasteleria` se
        # colaba por la puerta de atrás y machacaba in place el único producto
        # que ya está en v2.0. Las copias del scratchpad sí se pueden procesar:
        # son de usar y tirar y es donde se demuestra la idempotencia.
        dl_abs = os.path.abspath(DL)
        dentro_de_dl = ruta == dl_abs or ruta.startswith(dl_abs + os.sep)
        if dentro_de_dl and pid in EXCLUIDOS:
            print('ERROR: {} está excluido de la Fase A (ruta dentro de /dl)'.format(pid),
                  file=sys.stderr)
            return 1
        objetivos.append((pid, ruta))
    else:
        pid = args.objetivo
        if pid in EXCLUIDOS:
            print('ERROR: {} está excluido de la Fase A'.format(pid), file=sys.stderr)
            return 1
        ruta = os.path.join(DL, pid)
        if not os.path.isdir(ruta):
            print('ERROR: no existe {}'.format(ruta), file=sys.stderr)
            return 1
        objetivos.append((pid, ruta))

    informes = []
    todo_ok = True
    for pid, origen in objetivos:
        if args.dry_run:
            destino = os.path.join(SCRATCH, 'dryrun', pid)
            if os.path.isdir(destino):
                shutil.rmtree(destino)
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            shutil.copytree(origen, destino)
            carpeta = destino
        else:
            carpeta = origen
        corto = cortos.get(pid) or pid.replace('-', ' ').capitalize()
        inf = procesar_producto(carpeta, pid, corto, args.skip_cache)
        inf['dry_run'] = args.dry_run
        informes.append(inf)
        todo_ok &= inf['ok']
        print('{} {} → {} · patrones {} · cambios {}'.format(
            'OK  ' if inf['ok'] else 'FAIL', pid, carpeta,
            inf['patrones'], inf['total_cambios']))
        if inf['sin_contador']:
            print('   {} hoja(s) P2/P3/P4 sin contador:'.format(len(inf['sin_contador'])))
            for linea in inf['sin_contador']:
                print('     · {}'.format(linea))
        for f, d in inf['ficheros'].items():
            v = d['verificacion']
            if not v['ok']:
                print('   FALLA {}: {}'.format(f, {
                    k: v[k] for k in ('abre', 'creator', 'nocache_real', 'nonlat',
                                      'bio_vieja', 'empty_str', 'noprint',
                                      'contadores_ko')}))

    # Resumen final: lo que NO tumba la verificación pero hay que poder leer sin
    # abrir el JSON — hojas sin contador (INT-09) y fórmulas que pycel no sabe
    # evaluar (IRR/PMT/COUNTA), que se informan con hoja y celda y punto.
    total_sin_contador = sum(len(i['sin_contador']) for i in informes)
    pycel = [(i['producto'], f, c) for i in informes
             for f, d in i['ficheros'].items()
             for c in d['verificacion']['nocache_pycel']]
    print('— {} hoja(s) P2/P3/P4 sin contador · {} fórmula(s) sin cache por '
          'limitación de pycel'.format(total_sin_contador, len(pycel)))
    for prod, f, c in pycel:
        print('     · {}/{}:{}'.format(prod, f, c))

    if args.json_out:
        with open(args.json_out, 'w', encoding='utf-8') as fh:
            json.dump(informes if len(informes) > 1 else informes[0], fh,
                      ensure_ascii=False, indent=1)
    return 0 if todo_ok else 1


if __name__ == '__main__':
    sys.exit(main())
