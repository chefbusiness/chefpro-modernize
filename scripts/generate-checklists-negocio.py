#!/usr/bin/env python3
"""
Generate 24 Excel files (2 per kit × 12 kits) for restaurant business management checklists.
- File 1: Apertura/Cierre del Negocio
- File 2: Apertura/Cierre de Caja
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Brand Colors ──
GOLD = "FFD700"
DARK_BG = "1A1A1A"
HEADER_BG = "2D2D2D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"

# ── Base path ──
BASE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "public", "dl")

# ── Kit definitions ──
KITS = [
    {
        "dir": "kit-tareas",
        "name": "Kit de Tareas — Restaurante Casual",
        "concept": "Restaurante Casual",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [],
        "extra_cierre": [],
    },
    {
        "dir": "kit-tareas-cafeteria",
        "name": "Kit de Tareas — Cafetería / Brunch",
        "concept": "Cafetería / Brunch",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender máquinas de café y molinillo", "Barra"),
            ("Preparar barra de autoservicio", "Barra"),
        ],
        "extra_cierre": [
            ("Purgar y limpiar grupo de café", "Barra"),
        ],
    },
    {
        "dir": "kit-tareas-pizzeria",
        "name": "Kit de Tareas — Pizzería",
        "concept": "Pizzería",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender horno de leña (precalentar 1h antes del servicio)", "Cocina"),
            ("Montar barra de pizza si es visible", "Sala"),
        ],
        "extra_cierre": [
            ("Apagar horno de leña / dejar enfriar", "Cocina"),
        ],
    },
    {
        "dir": "kit-tareas-hamburgueseria",
        "name": "Kit de Tareas — Hamburguesería",
        "concept": "Hamburguesería",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender planchas y freidoras (precalentar 20 min)", "Cocina"),
            ("Montar línea de toppings", "Cocina"),
        ],
        "extra_cierre": [
            ("Limpiar planchas y filtrar aceite de freidoras", "Cocina"),
        ],
    },
    {
        "dir": "kit-tareas-dark-kitchen",
        "name": "Kit de Tareas — Dark Kitchen",
        "concept": "Dark Kitchen",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender tablets de plataformas delivery (Glovo, UberEats, JustEat)", "Operaciones"),
            ("Verificar stock de packaging", "Almacén"),
            ("Activar restaurante en plataformas", "Operaciones"),
        ],
        "extra_cierre": [
            ("Desactivar restaurante en plataformas delivery", "Operaciones"),
            ("Apagar tablets de plataformas", "Operaciones"),
        ],
    },
    {
        "dir": "kit-tareas-pasteleria",
        "name": "Kit de Tareas — Pastelería / Obrador",
        "concept": "Pastelería / Obrador",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender vitrinas refrigeradas de exposición", "Tienda"),
            ("Montar expositor de productos del día", "Tienda"),
        ],
        "extra_cierre": [
            ("Apagar vitrinas de exposición", "Tienda"),
            ("Retirar productos perecederos del expositor", "Tienda"),
        ],
    },
    {
        "dir": "kit-tareas-bar",
        "name": "Kit de Tareas — Bar / Cocktails",
        "concept": "Bar / Cocktails",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Verificar grifos de cerveza y presión CO2", "Barra"),
            ("Montar garnish station", "Barra"),
            ("Comprobar stock de hielo", "Barra"),
        ],
        "extra_cierre": [
            ("Limpiar grifos de cerveza", "Barra"),
            ("Vaciar y limpiar garnish station", "Barra"),
        ],
    },
    {
        "dir": "kit-tareas-catering",
        "name": "Kit de Tareas — Catering / Eventos",
        "concept": "Catering / Eventos",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Verificar vehículos de transporte", "Logística"),
            ("Comprobar equipamiento portátil", "Almacén"),
            ("Revisar orden de eventos del día", "Oficina"),
        ],
        "extra_cierre": [
            ("Descargar y limpiar vehículos de transporte", "Logística"),
            ("Inventariar equipamiento portátil devuelto", "Almacén"),
        ],
    },
    {
        "dir": "kit-tareas-hotel",
        "name": "Kit de Tareas — Hotel Completo",
        "concept": "Hotel Completo",
        "file1": "18-apertura-cierre-negocio.xlsx",
        "file2": "19-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Verificar check-ins del día", "Recepción"),
            ("Comprobar estado lobby y recepción", "Recepción"),
            ("Encender iluminación zonas nobles", "Zonas comunes"),
        ],
        "extra_cierre": [
            ("Verificar registro de check-outs pendientes", "Recepción"),
            ("Apagar iluminación de zonas nobles", "Zonas comunes"),
        ],
    },
    {
        "dir": "kit-tareas-heladeria",
        "name": "Kit de Tareas — Heladería Artesanal",
        "concept": "Heladería Artesanal",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender vitrinas/mantecadoras", "Tienda"),
            ("Comprobar temperaturas de todas las vitrinas", "Tienda"),
            ("Montar señalización de sabores", "Tienda"),
        ],
        "extra_cierre": [
            ("Cubrir vitrinas de helado", "Tienda"),
            ("Registrar temperaturas de cierre", "Tienda"),
        ],
    },
    {
        "dir": "kit-tareas-chocolateria",
        "name": "Kit de Tareas — Chocolatería / Bombonería",
        "concept": "Chocolatería / Bombonería",
        "file1": "08-apertura-cierre-negocio.xlsx",
        "file2": "09-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Encender vitrinas temperadas", "Tienda"),
            ("Verificar temperatura obrador (18-20°C)", "Obrador"),
        ],
        "extra_cierre": [
            ("Apagar vitrinas temperadas", "Tienda"),
            ("Registrar temperatura de cierre del obrador", "Obrador"),
        ],
    },
    {
        "dir": "kit-tareas-restaurante-creativo",
        "name": "Kit de Tareas — Restaurante Creativo / De Autor",
        "concept": "Restaurante Creativo / De Autor",
        "file1": "10-apertura-cierre-negocio.xlsx",
        "file2": "11-apertura-cierre-caja.xlsx",
        "extra_apertura": [
            ("Comprobar mise en place de sala (flores, decoración)", "Sala"),
            ("Verificar carta de vinos actualizada", "Sala"),
            ("Briefing de sala: menú degustación del día", "Sala"),
        ],
        "extra_cierre": [
            ("Retirar decoración fresca de mesas", "Sala"),
            ("Actualizar carta de vinos si hubo cambios", "Sala"),
        ],
    },
]

# ── Common tasks ──
COMMON_APERTURA = [
    ("Desactivar alarma del local", "Acceso"),
    ("Encender luces generales (interior y exterior)", "General"),
    ("Encender climatización (HVAC) / calefacción según temporada", "General"),
    ("Abrir persianas / puerta principal", "Acceso"),
    ("Montar señalización exterior / pizarra de menú", "Exterior"),
    ("Encender TPV / POS / datáfono", "Caja"),
    ("Encender sistema de reservas / tablet de pedidos", "Recepción"),
    ("Poner música ambiente al volumen adecuado", "General"),
    ("Revisar reservas del día y eventos especiales", "Recepción"),
    ("Comprobar baños: papel, jabón, limpieza, ambientador", "Baños"),
    ("Verificar limpieza de zonas comunes (entrada, pasillos)", "Zonas comunes"),
    ("Revisar estado del mobiliario (sillas, mesas, iluminación)", "Sala"),
    ("Montar terraza si aplica (mesas, sillas, sombrillas/estufas)", "Terraza"),
    ("Comprobar que las cartas/menús están en orden", "Sala"),
    ("Encender displays / pantallas informativas si las hay", "General"),
]

COMMON_CIERRE = [
    ("Apagar música ambiente", "General"),
    ("Verificar que no quedan clientes en el local", "Sala"),
    ("Limpiar y recoger terraza si aplica", "Terraza"),
    ("Apagar TPV / POS / datáfono (o dejar en standby)", "Caja"),
    ("Cerrar persianas y puertas", "Acceso"),
    ("Apagar luces (dejar mínimas de seguridad)", "General"),
    ("Apagar climatización", "General"),
    ("Recoger señalización exterior", "Exterior"),
    ("Comprobar que ventanas están cerradas", "General"),
    ("Vaciar papeleras de zonas comunes y baños", "Baños"),
    ("Revisión final de seguridad (gas, agua, enchufes)", "General"),
    ("Activar alarma", "Acceso"),
    ("Cerrar con llave y verificar cierre", "Acceso"),
]

# ── Styling helpers ──
THIN_BORDER = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

GOLD_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
DARK_FILL = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, bold=True, color=GOLD)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_BODY = Font(name="Calibri", size=10, color=DARK_BG)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=DARK_BG)
FONT_GOLD_HEADER = Font(name="Calibri", size=11, bold=True, color=DARK_BG)
FONT_BRAND = Font(name="Calibri", size=9, italic=True, color="888888")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def create_instructions_sheet(ws, product_name, file_description):
    """Create the Instrucciones tab."""
    ws.sheet_properties.tabColor = GOLD
    ws.column_dimensions["A"].width = 80

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    c = ws.cell(row=row, column=1, value="AI Chef Pro — aichef.pro")
    style_cell(c, FONT_TITLE, DARK_FILL, ALIGN_CENTER)
    ws.row_dimensions[row].height = 40

    row = 2
    c = ws.cell(row=row, column=1, value=product_name)
    style_cell(c, FONT_SUBTITLE, DARK_FILL, ALIGN_CENTER)
    ws.row_dimensions[row].height = 30

    row = 3
    c = ws.cell(row=row, column=1, value=file_description)
    style_cell(c, Font(name="Calibri", size=11, color=WHITE), DARK_FILL, ALIGN_CENTER)
    ws.row_dimensions[row].height = 25

    row = 5
    c = ws.cell(row=row, column=1, value="📋 Instrucciones de Uso")
    style_cell(c, Font(name="Calibri", size=14, bold=True, color=DARK_BG), GOLD_FILL, ALIGN_CENTER)
    ws.row_dimensions[row].height = 35

    instructions = [
        "1. Este archivo contiene checklists profesionales para la apertura y cierre de tu negocio.",
        "2. Imprime las hojas que necesites o úsalas en formato digital (tablet/ordenador).",
        "3. Cada checklist incluye columnas para marcar tareas completadas, responsable, hora y notas.",
        "4. Personaliza las tareas según las necesidades específicas de tu establecimiento.",
        "5. Añade o elimina filas según tu operativa diaria.",
        "6. Usa la validación de datos en la columna '✓ Completada' para marcar el estado.",
        "7. Rellena la firma y fecha cada día para mantener el control.",
        "",
        "💡 Consejo: Plastifica una copia impresa y usa rotuladores borrables para reutilizarla cada día.",
        "",
        "© AI Chef Pro — aichef.pro",
        "29 años de experiencia en alta hostelería · 15 años de consultoría gastronómica",
        "Contacto: info@aichef.pro",
    ]
    for i, text in enumerate(instructions):
        r = row + 1 + i
        c = ws.cell(row=r, column=1, value=text)
        style_cell(c, FONT_BODY, WHITE_FILL, ALIGN_LEFT)
        ws.row_dimensions[r].height = 22

    ws.print_area = "A1:A20"
    ws.sheet_format.defaultRowHeight = 18


def add_checklist_header_row(ws, row, columns):
    """Add header row for a checklist."""
    for col_idx, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=row, column=col_idx, value=label)
        style_cell(c, FONT_HEADER, HEADER_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 28


def add_checklist_data(ws, start_row, tasks, columns_count):
    """Add task rows. tasks = list of (tarea, zona)."""
    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Selecciona ✓, ✗ o —"
    dv.errorTitle = "Valor no válido"
    ws.add_data_validation(dv)

    for i, (tarea, zona) in enumerate(tasks):
        r = start_row + i
        fill = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
        row_height = 24

        # #
        c = ws.cell(row=r, column=1, value=i + 1)
        style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        # Tarea
        c = ws.cell(row=r, column=2, value=tarea)
        style_cell(c, FONT_BODY, fill, ALIGN_LEFT, THIN_BORDER)
        # Zona
        c = ws.cell(row=r, column=3, value=zona)
        style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        # Responsable
        c = ws.cell(row=r, column=4, value="")
        style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        # Hora Límite
        c = ws.cell(row=r, column=5, value="")
        style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        # ✓ Completada
        c = ws.cell(row=r, column=6, value="")
        style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        dv.add(c)
        # Firma
        c = ws.cell(row=r, column=7, value="")
        style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        # Notas
        c = ws.cell(row=r, column=8, value="")
        style_cell(c, FONT_BODY, fill, ALIGN_LEFT, THIN_BORDER)

        ws.row_dimensions[r].height = row_height

    return start_row + len(tasks)


def create_negocio_sheet(ws, title, tasks, concept_name):
    """Create an Apertura or Cierre del Negocio sheet."""
    ws.sheet_properties.tabColor = GOLD

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c = ws.cell(row=1, column=1, value=f"{title} — {concept_name}")
    style_cell(c, FONT_TITLE, DARK_FILL, ALIGN_CENTER)
    ws.row_dimensions[1].height = 40

    # Date / signature row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    c = ws.cell(row=2, column=1, value="Fecha: ____/____/________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)
    c = ws.cell(row=2, column=5, value="Responsable de turno: _________________________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.row_dimensions[2].height = 28

    # Empty row
    ws.row_dimensions[3].height = 8

    # Header
    columns = [
        ("#", 5),
        ("Tarea", 50),
        ("Zona", 16),
        ("Responsable", 16),
        ("Hora Límite", 13),
        ("✓ Completada", 14),
        ("Firma", 14),
        ("Notas", 25),
    ]
    add_checklist_header_row(ws, 4, columns)

    # Data
    next_row = add_checklist_data(ws, 5, tasks, 8)

    # Footer
    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)
    c = ws.cell(row=next_row, column=1, value="Firma del responsable: _________________________     Hora de finalización: ________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.row_dimensions[next_row].height = 30

    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)
    c = ws.cell(row=next_row, column=1, value="© AI Chef Pro — aichef.pro")
    style_cell(c, FONT_BRAND, None, ALIGN_RIGHT)

    # Print settings
    ws.print_area = f"A1:H{next_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_format.defaultRowHeight = 18


def generate_negocio_file(kit):
    """Generate the apertura/cierre del negocio file."""
    wb = Workbook()

    # Tab 1: Instrucciones
    ws_instr = wb.active
    ws_instr.title = "Instrucciones"
    create_instructions_sheet(
        ws_instr,
        kit["name"],
        "Checklist de Apertura y Cierre del Negocio",
    )

    # Tab 2: Apertura del Negocio
    ws_ap = wb.create_sheet("Apertura del Negocio")
    apertura_tasks = list(COMMON_APERTURA) + kit["extra_apertura"]
    create_negocio_sheet(ws_ap, "Apertura del Negocio", apertura_tasks, kit["concept"])

    # Tab 3: Cierre del Negocio
    ws_ci = wb.create_sheet("Cierre del Negocio")
    cierre_tasks = list(COMMON_CIERRE) + kit["extra_cierre"]
    create_negocio_sheet(ws_ci, "Cierre del Negocio", cierre_tasks, kit["concept"])

    path = os.path.join(BASE, kit["dir"], kit["file1"])
    wb.save(path)
    return path


# ── Caja tasks ──
APERTURA_CAJA = [
    "Verificar fondo de caja del día anterior",
    "Contar fondo de caja inicial",
    "Registrar importe de fondo de caja en hoja",
    "Verificar cambio suficiente (monedas, billetes pequeños)",
    "Encender TPV / POS",
    "Imprimir X de apertura (informe de inicio)",
    "Comprobar rollo de papel de ticket",
    "Verificar datáfono operativo",
    "Verificar conexión de pasarela de pago",
    "Anotar incidencias del turno anterior",
    "Firma del responsable de apertura",
]

CIERRE_CAJA = [
    "Imprimir Z de cierre (informe final del día)",
    "Recuento de efectivo por denominación (ver tabla inferior)",
    "Calcular total efectivo",
    "Anotar total tarjetas (Visa/Mastercard)",
    "Anotar otros medios (Bizum, vales, invitaciones)",
    "Total facturado = efectivo + tarjetas + otros",
    "Comparar con Z del TPV → calcular diferencia/descuadre",
    "Registrar devoluciones del día",
    "Registrar anulaciones del día",
    "Separar fondo de caja para mañana",
    "Preparar depósito bancario / guardar en caja fuerte",
    "Firma del responsable de cierre",
]

DENOMINACIONES_BILLETES = [
    ("500 €", 500),
    ("200 €", 200),
    ("100 €", 100),
    ("50 €", 50),
    ("20 €", 20),
    ("10 €", 10),
    ("5 €", 5),
]

DENOMINACIONES_MONEDAS = [
    ("2 €", 2),
    ("1 €", 1),
    ("0,50 €", 0.50),
    ("0,20 €", 0.20),
    ("0,10 €", 0.10),
    ("0,05 €", 0.05),
    ("0,01 €", 0.01),
]


def create_caja_checklist_sheet(ws, title, tasks, concept_name):
    """Create Apertura or Cierre de Caja checklist sheet."""
    ws.sheet_properties.tabColor = GOLD

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value=f"{title} — {concept_name}")
    style_cell(c, FONT_TITLE, DARK_FILL, ALIGN_CENTER)
    ws.row_dimensions[1].height = 40

    # Date/signature
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    c = ws.cell(row=2, column=1, value="Fecha: ____/____/________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=6)
    c = ws.cell(row=2, column=4, value="Responsable: _________________________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 8

    # Header
    columns = [
        ("#", 5),
        ("Tarea", 50),
        ("Responsable", 16),
        ("✓ Completada", 14),
        ("Hora", 12),
        ("Notas", 25),
    ]
    for col_idx, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=4, column=col_idx, value=label)
        style_cell(c, FONT_HEADER, HEADER_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 28

    # Data validation
    dv = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
    dv.error = "Selecciona ✓, ✗ o —"
    dv.errorTitle = "Valor no válido"
    ws.add_data_validation(dv)

    for i, tarea in enumerate(tasks):
        r = 5 + i
        fill = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row=r, column=1, value=i + 1)
        style_cell(ws.cell(row=r, column=1), FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        ws.cell(row=r, column=2, value=tarea)
        style_cell(ws.cell(row=r, column=2), FONT_BODY, fill, ALIGN_LEFT, THIN_BORDER)
        for col in range(3, 7):
            ws.cell(row=r, column=col, value="")
            style_cell(ws.cell(row=r, column=col), FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
        dv.add(ws.cell(row=r, column=4))
        ws.row_dimensions[r].height = 24

    next_row = 5 + len(tasks)

    # If this is Cierre, add cash counting table
    if "Cierre" in title:
        next_row += 1
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
        c = ws.cell(row=next_row, column=1, value="💰 Recuento de Efectivo por Denominación")
        style_cell(c, Font(name="Calibri", size=13, bold=True, color=DARK_BG), GOLD_FILL, ALIGN_CENTER)
        ws.row_dimensions[next_row].height = 32

        next_row += 1
        for col_idx, (label, width) in enumerate([("Denominación", 16), ("Cantidad", 14), ("Subtotal (€)", 16)], 1):
            c = ws.cell(row=next_row, column=col_idx, value=label)
            style_cell(c, FONT_HEADER, HEADER_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.row_dimensions[next_row].height = 26

        # Section: Billetes
        next_row += 1
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=3)
        c = ws.cell(row=next_row, column=1, value="BILLETES")
        style_cell(c, FONT_GOLD_HEADER, GOLD_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.row_dimensions[next_row].height = 24

        for denom_label, denom_val in DENOMINACIONES_BILLETES:
            next_row += 1
            fill = LIGHT_FILL if DENOMINACIONES_BILLETES.index((denom_label, denom_val)) % 2 == 0 else WHITE_FILL
            c = ws.cell(row=next_row, column=1, value=denom_label)
            style_cell(c, FONT_BODY_BOLD, fill, ALIGN_CENTER, THIN_BORDER)
            c = ws.cell(row=next_row, column=2, value=0)
            style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
            c = ws.cell(row=next_row, column=3)
            c.value = f"=B{next_row}*{denom_val}"
            c.number_format = '#,##0.00 €'
            style_cell(c, FONT_BODY, fill, ALIGN_RIGHT, THIN_BORDER)
            ws.row_dimensions[next_row].height = 22

        billetes_start = next_row - len(DENOMINACIONES_BILLETES) + 1

        # Section: Monedas
        next_row += 1
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=3)
        c = ws.cell(row=next_row, column=1, value="MONEDAS")
        style_cell(c, FONT_GOLD_HEADER, GOLD_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.row_dimensions[next_row].height = 24

        for denom_label, denom_val in DENOMINACIONES_MONEDAS:
            next_row += 1
            fill = LIGHT_FILL if DENOMINACIONES_MONEDAS.index((denom_label, denom_val)) % 2 == 0 else WHITE_FILL
            c = ws.cell(row=next_row, column=1, value=denom_label)
            style_cell(c, FONT_BODY_BOLD, fill, ALIGN_CENTER, THIN_BORDER)
            c = ws.cell(row=next_row, column=2, value=0)
            style_cell(c, FONT_BODY, fill, ALIGN_CENTER, THIN_BORDER)
            c = ws.cell(row=next_row, column=3)
            c.value = f"=B{next_row}*{denom_val}"
            c.number_format = '#,##0.00 €'
            style_cell(c, FONT_BODY, fill, ALIGN_RIGHT, THIN_BORDER)
            ws.row_dimensions[next_row].height = 22

        monedas_end = next_row

        # Total efectivo
        next_row += 1
        c = ws.cell(row=next_row, column=1, value="TOTAL EFECTIVO")
        style_cell(c, Font(name="Calibri", size=11, bold=True, color=WHITE), DARK_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
        c = ws.cell(row=next_row, column=3)
        c.value = f"=SUM(C{billetes_start}:C{monedas_end})"
        c.number_format = '#,##0.00 €'
        style_cell(c, Font(name="Calibri", size=12, bold=True, color=WHITE), DARK_FILL, ALIGN_RIGHT, THIN_BORDER)
        ws.row_dimensions[next_row].height = 28

        # Summary section
        next_row += 2
        summary_items = [
            ("Total Efectivo", f"=C{next_row - 2}"),
            ("Total Tarjetas (Visa/MC)", ""),
            ("Total Otros (Bizum, Vales)", ""),
            ("TOTAL FACTURADO", None),  # formula below
            ("Z del TPV", ""),
            ("DESCUADRE", None),  # formula below
        ]

        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=3)
        c = ws.cell(row=next_row, column=1, value="📊 Resumen de Cierre")
        style_cell(c, Font(name="Calibri", size=13, bold=True, color=DARK_BG), GOLD_FILL, ALIGN_CENTER)
        ws.row_dimensions[next_row].height = 30

        next_row += 1
        for col_idx, (label, width) in enumerate([("Concepto", 30), ("Importe (€)", 20)], 1):
            c = ws.cell(row=next_row, column=col_idx, value=label)
            style_cell(c, FONT_HEADER, HEADER_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.row_dimensions[next_row].height = 26

        summary_start = next_row + 1
        for i, (label, val) in enumerate(summary_items):
            r = summary_start + i
            fill = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
            is_total = label in ("TOTAL FACTURADO", "DESCUADRE")

            c = ws.cell(row=r, column=1, value=label)
            style_cell(c, FONT_BODY_BOLD if is_total else FONT_BODY, GOLD_FILL if is_total else fill, ALIGN_LEFT, THIN_BORDER)

            c = ws.cell(row=r, column=2)
            if label == "TOTAL FACTURADO":
                c.value = f"=B{summary_start}+B{summary_start+1}+B{summary_start+2}"
            elif label == "DESCUADRE":
                c.value = f"=B{summary_start+3}-B{summary_start+4}"
            else:
                c.value = val if val else 0
            c.number_format = '#,##0.00 €'
            style_cell(c, FONT_BODY_BOLD if is_total else FONT_BODY, GOLD_FILL if is_total else fill, ALIGN_RIGHT, THIN_BORDER)
            ws.row_dimensions[r].height = 24

        next_row = summary_start + len(summary_items)

    # Footer
    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
    c = ws.cell(row=next_row, column=1, value="Firma del responsable: _________________________     Hora: ________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.row_dimensions[next_row].height = 30

    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
    c = ws.cell(row=next_row, column=1, value="© AI Chef Pro — aichef.pro")
    style_cell(c, FONT_BRAND, None, ALIGN_RIGHT)

    ws.print_area = f"A1:F{next_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def create_registro_mensual_sheet(ws, concept_name):
    """Create Registro Mensual tab — same for all kits."""
    ws.sheet_properties.tabColor = GOLD

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c = ws.cell(row=1, column=1, value=f"Registro Mensual de Caja — {concept_name}")
    style_cell(c, FONT_TITLE, DARK_FILL, ALIGN_CENTER)
    ws.row_dimensions[1].height = 40

    # Month/Year
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    c = ws.cell(row=2, column=1, value="Mes: ________________     Año: ________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=9)
    c = ws.cell(row=2, column=6, value="Responsable general: _________________________")
    style_cell(c, FONT_BODY_BOLD, GOLD_FILL, ALIGN_LEFT)
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 8

    # Headers
    headers = [
        ("Fecha", 12),
        ("Fondo Apertura", 16),
        ("Ventas Efectivo", 16),
        ("Ventas Tarjeta", 16),
        ("Ventas Otros", 14),
        ("Total Facturado", 16),
        ("Descuadre", 14),
        ("Depósito", 14),
        ("Responsable", 16),
    ]
    for col_idx, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=label)
        style_cell(c, FONT_HEADER, HEADER_FILL, ALIGN_CENTER, THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 28

    # 31 rows
    for day in range(1, 32):
        r = 4 + day
        fill = LIGHT_FILL if day % 2 == 1 else WHITE_FILL
        c = ws.cell(row=r, column=1, value=day)
        style_cell(c, FONT_BODY_BOLD, fill, ALIGN_CENTER, THIN_BORDER)
        for col in range(2, 10):
            c = ws.cell(row=r, column=col, value="" if col == 9 else 0)
            if col == 6:
                c.value = f"=C{r}+D{r}+E{r}"
            if col == 7:
                c.value = f"=F{r}-B{r}" if col == 7 else c.value
                # descuadre = total facturado minus something? Let's define:
                # Descuadre is typically difference between POS Z and actual count
                # Keep it editable
                c.value = ""
            if col in (2, 3, 4, 5, 8):
                c.value = 0 if col != 9 else ""
            if col == 6:
                c.value = f"=C{r}+D{r}+E{r}"
                c.number_format = '#,##0.00 €'
            elif col in (2, 3, 4, 5, 7, 8):
                c.number_format = '#,##0.00 €'
            style_cell(c, FONT_BODY, fill, ALIGN_CENTER if col == 9 else ALIGN_RIGHT, THIN_BORDER)
        ws.row_dimensions[r].height = 22

    # Totals row
    total_row = 36
    c = ws.cell(row=total_row, column=1, value="TOTALES")
    style_cell(c, Font(name="Calibri", size=11, bold=True, color=WHITE), DARK_FILL, ALIGN_CENTER, THIN_BORDER)

    for col in range(2, 9):
        c = ws.cell(row=total_row, column=col)
        col_letter = get_column_letter(col)
        c.value = f"=SUM({col_letter}5:{col_letter}35)"
        c.number_format = '#,##0.00 €'
        style_cell(c, Font(name="Calibri", size=11, bold=True, color=WHITE), DARK_FILL, ALIGN_RIGHT, THIN_BORDER)

    c = ws.cell(row=total_row, column=9, value="")
    style_cell(c, Font(name="Calibri", size=11, bold=True, color=WHITE), DARK_FILL, ALIGN_CENTER, THIN_BORDER)
    ws.row_dimensions[total_row].height = 28

    # Footer
    r = total_row + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value="© AI Chef Pro — aichef.pro")
    style_cell(c, FONT_BRAND, None, ALIGN_RIGHT)

    ws.print_area = f"A1:I{r}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def generate_caja_file(kit):
    """Generate the apertura/cierre de caja file."""
    wb = Workbook()

    # Tab 1: Instrucciones
    ws_instr = wb.active
    ws_instr.title = "Instrucciones"
    create_instructions_sheet(
        ws_instr,
        kit["name"],
        "Checklist de Apertura y Cierre de Caja",
    )

    # Tab 2: Apertura de Caja
    ws_ap = wb.create_sheet("Apertura de Caja")
    create_caja_checklist_sheet(ws_ap, "Apertura de Caja", APERTURA_CAJA, kit["concept"])

    # Tab 3: Cierre de Caja
    ws_ci = wb.create_sheet("Cierre de Caja")
    create_caja_checklist_sheet(ws_ci, "Cierre de Caja", CIERRE_CAJA, kit["concept"])

    # Tab 4: Registro Mensual
    ws_reg = wb.create_sheet("Registro Mensual")
    create_registro_mensual_sheet(ws_reg, kit["concept"])

    path = os.path.join(BASE, kit["dir"], kit["file2"])
    wb.save(path)
    return path


# ── Main ──
if __name__ == "__main__":
    print("=" * 60)
    print("Generando checklists de Apertura/Cierre de Negocio y Caja")
    print("=" * 60)

    created = []
    for kit in KITS:
        dir_path = os.path.join(BASE, kit["dir"])
        if not os.path.isdir(dir_path):
            print(f"⚠️  Directorio no encontrado: {dir_path}")
            continue

        f1 = generate_negocio_file(kit)
        print(f"✅ {kit['concept']:40s} → {os.path.basename(f1)}")
        created.append(f1)

        f2 = generate_caja_file(kit)
        print(f"✅ {kit['concept']:40s} → {os.path.basename(f2)}")
        created.append(f2)

    print()
    print(f"Total archivos creados: {len(created)} / 24")
    print("=" * 60)
    for f in created:
        rel = os.path.relpath(f, os.path.join(BASE, ".."))
        print(f"  {rel}")
